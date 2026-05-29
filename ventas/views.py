from __future__ import annotations

from collections import OrderedDict, defaultdict
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from celery.result import AsyncResult
from django.core.exceptions import PermissionDenied
from django.db.models import Count, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.access import ROLE_ADMIN, ROLE_COMPRAS, ROLE_DG, ROLE_PRODUCCION, has_any_role
from core.models import Sucursal
from pos_bridge.models import PointProduct, PointSalesDailyProductFact
from ventas.models import PronosticoGuardado
from ventas.services.pronostico_engine import (
    MONTHS_ES,
    ORDEN_CATEGORIAS,
    WEEKDAYS_ES,
)
from ventas.services.pronostico_engine import calcular_pronostico
from ventas.services.sales_freshness import (
    get_forecast_sales_freshness,
    queue_forecast_sales_refresh_if_needed,
)
from ventas.tasks import calcular_y_guardar_pronostico


EXCLUDED_PRODUCT_CATEGORIES = {
    "COCA-COLA",
    "CLARITA",
    "HIELO Y AGUA MAR DE CORTÉZ",
    "INDUSTRIAS LEC",
    "MEDIA PLANCHA",
    "PLÁSTICOS",
    "PLASTICOS",
    "REGALOS",
    "ROSCA",
    "SAN VALENTÍN",
    "SAN VALENTIN",
    "TE",
}
EXCLUDED_PRODUCT_NAME_TERMS = (
    "tarjeta de regalo",
    "servicio domicilio",
    "extra 100",
    "media plancha",
)
ACCESSORY_CATEGORIES = {
    "ACCESORIOS DE REPOSTERÍA",
    "ALEGRIA",
    "ALEGRÍA",
    "GLOW",
    "GRANMARK",
    "PILLINES",
    "VELAS SPARKLERS",
    "VIVA PARTY",
}
BEVERAGE_CATEGORIES = {
    "CAFE",
    "CAFÉ",
}
CATALOG_CATEGORY_ORDER = [
    "Bollo",
    "Empanadas",
    "Galletas",
    "Cheesecake",
    "Individual",
    "Pastel Grande",
    "Pastel Mediano",
    "Pastel Chico",
    "Pastel Mini",
    "Rebanada",
    "Pay Grande",
    "Pay Mediano",
    "Vasos Preparados Grande",
    "Otros postres",
    "Bebidas",
    "Accesorios",
]
CATALOG_CATEGORY_INDEX = {category.casefold(): index for index, category in enumerate(CATALOG_CATEGORY_ORDER)}


def _is_excluded_product_category(category: str) -> bool:
    normalized = _clean_category_label(category).upper()
    if normalized in EXCLUDED_PRODUCT_CATEGORIES:
        return True
    return False


def _is_excluded_product_name(name: str) -> bool:
    name_fold = (name or "").casefold()
    return any(term in name_fold for term in EXCLUDED_PRODUCT_NAME_TERMS)


def _clean_category_label(value: str | None) -> str:
    return " ".join((value or "Sin categoría").strip().split()) or "Sin categoría"


def _catalog_category_sort_key(category: str) -> tuple[int, int, str]:
    label = _clean_category_label(category)
    index = CATALOG_CATEGORY_INDEX.get(label.casefold())
    if index is not None:
        return (0, index, label.casefold())
    return (1, len(CATALOG_CATEGORY_ORDER), label.casefold())


def _category_for_catalog_product(product: PointProduct) -> str:
    category = _clean_category_label(product.category)
    name = product.name or ""
    category_upper = category.upper()
    if category_upper in ACCESSORY_CATEGORIES:
        return "Accesorios"
    if category_upper in BEVERAGE_CATEGORIES:
        return "Bebidas"
    if category_upper in {"VASOS GRANDE", "VASOS MINI", "VASO PREPARADO MINI"}:
        return "Vasos Preparados Grande"
    is_sabor = "sabor" in category.casefold() or name.casefold().startswith("sabor")
    if not is_sabor:
        return category
    name_fold = name.casefold()
    if "grande" in name_fold:
        return "Pay Grande"
    if "mediano" in name_fold:
        return "Pay Mediano"
    if "rebanada" in name_fold:
        return "Rebanada"
    if "chico" in name_fold:
        return "Pay Chico"
    if "mini" in name_fold:
        return "Pay Mini"
    return category


def _can_view_pronostico(user) -> bool:
    return has_any_role(user, ROLE_DG, ROLE_ADMIN, ROLE_PRODUCCION, ROLE_COMPRAS, "VENTAS", "LECTURA")


def _json_ready(value):
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, defaultdict):
        return {str(key): _json_ready(val) for key, val in value.items()}
    if isinstance(value, dict):
        return {str(key): _json_ready(val) for key, val in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_ready(item) for item in value]
    return value


def _decimal_from_json(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _int_from_json(value) -> int:
    if value in (None, ""):
        return 0
    return int(Decimal(str(value)))


def _parse_iso_date(value: str) -> date | None:
    try:
        return date.fromisoformat(str(value))
    except (TypeError, ValueError):
        return None


def _date_header(value: str) -> str:
    parsed = _parse_iso_date(value)
    if not parsed:
        return str(value)
    return f"{WEEKDAYS_ES[parsed.weekday()]} {parsed.day} {MONTHS_ES[parsed.month - 1]}"


def _selected_branch_ids(request, *, source: str = "GET") -> set[int]:
    values = request.GET.getlist("sucursales") if source == "GET" else request.POST.getlist("sucursales")
    selected = {int(value) for value in values if str(value).isdigit()}
    active = set(Sucursal.objects.filter(activa=True).values_list("id", flat=True))
    return (selected & active) if selected else set(active)


def _selected_product_skus(request) -> list[str]:
    return [value.strip() for value in request.POST.getlist("productos_incluidos") if value.strip()]


def _catalogo_productos_por_categoria() -> OrderedDict[str, list[dict]]:
    hace_30 = timezone.localdate() - timedelta(days=30)
    skus_vigentes = set(
        PointSalesDailyProductFact.objects.filter(
            sale_date__gte=hace_30,
            point_product__active=True,
            point_product__sku__gt="",
        )
        .exclude(point_product__name__istartswith="TOPPING")
        .exclude(point_product__name__icontains="topping")
        .exclude(point_product__name__icontains="tarjeta de regalo")
        .exclude(point_product__name__icontains="servicio domicilio")
        .exclude(point_product__name__icontains="extra 100")
        .exclude(point_product__name__icontains="media plancha")
        .values_list("point_product__sku", flat=True)
        .distinct()
        )

    if not skus_vigentes:
        # Si no hay ventas recientes, mostramos el catálogo activo completo para no bloquear el pronóstico.
        skus_vigentes = set(
            PointProduct.objects.filter(active=True, sku__gt="")
            .exclude(name__istartswith="TOPPING")
            .exclude(name__icontains="topping")
            .exclude(name__icontains="tarjeta de regalo")
            .exclude(name__icontains="servicio domicilio")
            .exclude(name__icontains="extra 100")
            .exclude(name__icontains="media plancha")
            .values_list("sku", flat=True)
            .distinct()
        )
    pay_durazno_skus = set(
        PointSalesDailyProductFact.objects.filter(
            sale_date__gte=timezone.localdate() - timedelta(days=90),
            point_product__active=True,
            point_product__name__icontains="Pay de Queso con Durazno",
        )
        .values("point_product__sku")
        .annotate(dias=Count("sale_date", distinct=True))
        .filter(dias__lte=3)
        .values_list("point_product__sku", flat=True)
    )
    categorias_raw: dict[str, list[dict]] = defaultdict(list)
    products = (
        PointProduct.objects.filter(active=True, sku__in=skus_vigentes)
        .exclude(name__istartswith="TOPPING")
        .exclude(name__icontains="topping")
        .only("id", "sku", "name", "category")
        .order_by("category", "name")
    )
    for product in products:
        if _is_excluded_product_name(product.name):
            continue
        if product.sku in pay_durazno_skus:
            continue
        category = _category_for_catalog_product(product)
        if _is_excluded_product_category(category):
            continue
        categorias_raw[category].append(
            {
                "id": product.id,
                "nombre": product.name,
                "sku": product.sku,
            }
        )

    categorias = OrderedDict()
    for category in sorted(categorias_raw, key=_catalog_category_sort_key):
        categorias[category] = sorted(categorias_raw[category], key=lambda product: product["nombre"].casefold())
    return categorias


def _build_comparativa_real(pronostico: PronosticoGuardado, resultados: dict) -> tuple[bool, list[dict], dict]:
    if pronostico.fecha_fin >= timezone.localdate():
        return False, [], {}

    productos = []
    for category in resultados.get("por_categoria") or []:
        for product in category.get("productos") or []:
            product_id = product.get("point_product_id")
            if not product_id:
                continue
            productos.append(
                {
                    "point_product_id": int(product_id),
                    "nombre": product.get("nombre") or "Producto",
                    "categoria": product.get("categoria") or category.get("categoria") or "Sin categoría",
                    "pronostico_piezas": _decimal_from_json(product.get("total_piezas")),
                }
            )
    if not productos:
        return False, [], {}

    product_ids = {item["point_product_id"] for item in productos}
    real_qs = PointSalesDailyProductFact.objects.filter(
        sale_date__range=(pronostico.fecha_inicio, pronostico.fecha_fin),
        point_product_id__in=product_ids,
    )
    branch_ids = set(pronostico.sucursales.values_list("id", flat=True))
    if branch_ids:
        real_qs = real_qs.filter(branch__erp_branch_id__in=branch_ids)

    reales = {
        int(row["point_product_id"]): _decimal_from_json(row["real_piezas"])
        for row in real_qs.values("point_product_id").annotate(real_piezas=Sum("total_cantidad"))
        if row.get("point_product_id")
    }
    if not reales:
        return False, [], {}

    comparativa = []
    total_pronostico = Decimal("0")
    total_real = Decimal("0")
    productos_acertados = 0
    productos_sobre = 0
    productos_bajo = 0
    for product in productos:
        pron = product["pronostico_piezas"]
        real = reales.get(product["point_product_id"], Decimal("0"))
        diferencia = real - pron
        diferencia_pct = (diferencia / pron * Decimal("100")) if pron > 0 else Decimal("0")
        abs_pct = abs(diferencia_pct)
        if abs_pct < Decimal("15"):
            alerta = "ok"
            productos_acertados += 1
        elif real > pron * Decimal("1.15"):
            alerta = "sobre"
            productos_sobre += 1
        else:
            alerta = "bajo"
            productos_bajo += 1

        total_pronostico += pron
        total_real += real
        comparativa.append(
            {
                "nombre": product["nombre"],
                "categoria": product["categoria"],
                "pronostico_piezas": pron,
                "real_piezas": real,
                "diferencia": diferencia,
                "diferencia_pct": diferencia_pct,
                "alerta": alerta,
                "_abs_pct": abs_pct,
            }
        )

    comparativa.sort(key=lambda item: (-item["_abs_pct"], item["categoria"], item["nombre"]))
    for item in comparativa:
        item.pop("_abs_pct", None)
    diferencia_total = total_real - total_pronostico
    precision_pct = Decimal("0")
    if total_real > 0:
        precision_pct = max(Decimal("0"), Decimal("100") - (abs(diferencia_total) / total_real * Decimal("100")))
    resumen_comparativa = {
        "total_pronostico": total_pronostico,
        "total_real": total_real,
        "diferencia_total": diferencia_total,
        "precision_pct": precision_pct,
        "productos_acertados": productos_acertados,
        "productos_sobre": productos_sobre,
        "productos_bajo": productos_bajo,
    }
    return True, comparativa, resumen_comparativa


def _validate_dates(fecha_inicio_raw: str, fecha_fin_raw: str) -> tuple[date | None, date | None, list[str]]:
    errors = []
    fecha_inicio = _parse_iso_date(fecha_inicio_raw)
    fecha_fin = _parse_iso_date(fecha_fin_raw)
    if not fecha_inicio or not fecha_fin:
        errors.append("Selecciona fechas validas para calcular el pronostico.")
        return fecha_inicio, fecha_fin, errors
    if fecha_inicio > fecha_fin:
        errors.append("La fecha inicio no puede ser posterior a la fecha fin.")
    if (fecha_fin - fecha_inicio).days > 45:
        errors.append("El rango maximo permitido es de 46 dias para mantener el calculo operativo.")
    return fecha_inicio, fecha_fin, errors


def _empty_result_context(resultados: dict | None) -> dict:
    resultados = resultados or {}
    resumen = resultados.get("resumen")
    if resumen:
        resumen = dict(resumen)
        resumen.setdefault("n_productos", resumen.get("productos", 0))
        resumen.setdefault("n_sucursales", resumen.get("sucursales", 0))
        resumen["confianza_pct"] = float(_decimal_from_json(resumen.get("confianza_promedio"))) * 100
    return {
        "resultados": resultados,
        "resumen": resumen,
        "fechas_tabla": resultados.get("fechas_tabla") or [],
        "fechas_especiales_en_rango": (resultados.get("resumen") or {}).get("fechas_especiales") or [],
        "por_categoria": resultados.get("por_categoria") or [],
        "por_dia": resultados.get("por_dia") or [],
        "por_sucursal": resultados.get("por_sucursal") or [],
    }


def _safe_sheet_title(raw_title: str, used_titles: set[str]) -> str:
    invalid_chars = '[]:*?/\\'
    title = "".join("-" if char in invalid_chars else char for char in (raw_title or "Sucursal")).strip()
    title = title[:31] or "Sucursal"
    base_title = title
    counter = 2
    while title in used_titles:
        suffix = f" {counter}"
        title = f"{base_title[:31 - len(suffix)]}{suffix}"
        counter += 1
    used_titles.add(title)
    return title


def _style_row(ws, row_number: int, *, fill: str | None = None, font_color: str = "000000", bold: bool = False):
    row_fill = PatternFill("solid", fgColor=fill) if fill else None
    row_font = Font(color=font_color, bold=bold)
    for cell in ws[row_number]:
        if row_fill:
            cell.fill = row_fill
        cell.font = row_font
        cell.alignment = Alignment(vertical="center")


def _product_day_value(product: dict, fecha_iso: str, escenario: str = "recomendado") -> int:
    dias = product.get("dias") or {}
    day_data = dias.get(fecha_iso) or {}
    return _int_from_json(day_data.get(escenario, 0))


def _write_pronostico_sheet(ws, *, title: str, subtitle: str, fechas: list[str], categorias: list[dict]):
    ws["A1"] = title
    ws["A1"].font = Font(color="7B1A48", bold=True, size=14)
    ws["A2"] = subtitle
    ws["A2"].font = Font(color="555555", size=10)
    ws.cell(row=3, column=1, value="")

    headers = ["Categoría", "Producto"] + [_date_header(fecha) for fecha in fechas] + ["Total", "Precio", "Ingreso"]
    ws.append(headers)
    _style_row(ws, 4, fill="F5E6ED", font_color="7B1A48", bold=True)
    ws.freeze_panes = "A5"

    data_start_row = 5
    current_row = data_start_row
    grand_total_by_day = {fecha: 0 for fecha in fechas}
    grand_total_pieces = 0
    grand_total_income = Decimal("0")

    category_map = {category.get("categoria"): category for category in categorias}
    for category_name in ORDEN_CATEGORIAS:
        category = category_map.get(category_name)
        if not category:
            continue

        subtotal_by_day = {fecha: 0 for fecha in fechas}
        subtotal_pieces = 0
        subtotal_income = Decimal("0")
        for product in category.get("productos") or []:
            total_pieces = _int_from_json(product.get("total_piezas"))
            if total_pieces <= 0:
                continue
            day_values = [_product_day_value(product, fecha) for fecha in fechas]
            price = _decimal_from_json(product.get("precio"))
            income = _decimal_from_json(product.get("total_ingreso"))
            ws.append([category_name, product.get("nombre") or "Producto"] + day_values + [total_pieces, float(price), float(income)])
            if (current_row - data_start_row) % 2:
                _style_row(ws, current_row, fill="FDF2F6")

            for fecha, value in zip(fechas, day_values, strict=False):
                subtotal_by_day[fecha] += value
                grand_total_by_day[fecha] += value
            subtotal_pieces += total_pieces
            subtotal_income += income
            grand_total_pieces += total_pieces
            grand_total_income += income
            current_row += 1

        ws.append([f"Subtotal {category_name}", ""] + [subtotal_by_day[fecha] for fecha in fechas] + [subtotal_pieces, "", float(subtotal_income)])
        _style_row(ws, current_row, fill="7B1A48", font_color="FFFFFF", bold=True)
        current_row += 1

    ws.append(["Total general", ""] + [grand_total_by_day[fecha] for fecha in fechas] + [grand_total_pieces, "", float(grand_total_income)])
    _style_row(ws, current_row, fill="3D0A24", font_color="FFFFFF", bold=True)

    first_day_col = 3
    total_col = first_day_col + len(fechas)
    price_col = total_col + 1
    income_col = price_col + 1
    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row):
        for index, cell in enumerate(row, start=1):
            if first_day_col <= index <= total_col:
                cell.number_format = "#,##0"
            elif index == price_col:
                cell.number_format = "$#,##0.00"
            elif index == income_col:
                cell.number_format = "$#,##0"

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 44)


def _write_escenarios_sheet(ws, *, title: str, subtitle: str, categorias: list[dict]):
    ws["A1"] = title
    ws["A1"].font = Font(color="7B1A48", bold=True, size=14)
    ws["A2"] = subtitle
    ws["A2"].font = Font(color="555555", size=10)
    ws.cell(row=3, column=1, value="")
    ws.append(["Categoría", "Producto", "Conservador", "Recomendado", "Agresivo", "Precio", "Ingreso recomendado"])
    _style_row(ws, 4, fill="F5E6ED", font_color="7B1A48", bold=True)
    ws.freeze_panes = "A5"

    current_row = 5
    category_map = {category.get("categoria"): category for category in categorias}
    for category_name in ORDEN_CATEGORIAS:
        category = category_map.get(category_name)
        if not category:
            continue
        for product in category.get("productos") or []:
            escenarios = product.get("escenarios") or {}
            price = _decimal_from_json(product.get("precio"))
            income = _decimal_from_json(product.get("total_ingreso"))
            ws.append(
                [
                    category_name,
                    product.get("nombre") or "Producto",
                    _int_from_json(escenarios.get("conservador")),
                    _int_from_json(escenarios.get("recomendado") or product.get("total_piezas")),
                    _int_from_json(escenarios.get("agresivo")),
                    float(price),
                    float(income),
                ]
            )
            if current_row % 2:
                _style_row(ws, current_row, fill="FDF2F6")
            current_row += 1

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for index, cell in enumerate(row, start=1):
            if index in {3, 4, 5}:
                cell.number_format = "#,##0"
            elif index in {6, 7}:
                cell.number_format = "$#,##0.00"

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 44)


def _build_pronostico_excel_response(pronostico: PronosticoGuardado) -> HttpResponse:
    resultados = pronostico.resultado_json or {}
    fechas = resultados.get("fechas") or []
    resumen = resultados.get("resumen") or {}
    workbook = Workbook()
    used_titles = set()
    generated_at = timezone.localtime(pronostico.creado_en).strftime("%Y-%m-%d %H:%M")
    user_label = pronostico.creado_por.get_full_name() if pronostico.creado_por else ""
    user_label = user_label or (pronostico.creado_por.get_username() if pronostico.creado_por else "Sin usuario")
    subtitle = (
        f"Metodo: {resumen.get('metodo', 'pronostico')} | "
        f"Rango: {pronostico.fecha_inicio} a {pronostico.fecha_fin} | "
        f"Generado: {generated_at} | Por: {user_label}"
    )

    summary_sheet = workbook.active
    summary_sheet.title = _safe_sheet_title("Resumen general", used_titles)
    _write_pronostico_sheet(
        summary_sheet,
        title=pronostico.nombre,
        subtitle=subtitle,
        fechas=fechas,
        categorias=resultados.get("por_categoria") or [],
    )

    for branch in resultados.get("por_sucursal") or []:
        sheet = workbook.create_sheet(_safe_sheet_title(branch.get("sucursal_nombre") or branch.get("sucursal") or "Sucursal", used_titles))
        _write_pronostico_sheet(
            sheet,
            title=f"{pronostico.nombre} - {branch.get('sucursal_nombre') or branch.get('sucursal') or 'Sucursal'}",
            subtitle=subtitle,
            fechas=fechas,
            categorias=branch.get("categorias") or [],
        )

    scenarios_sheet = workbook.create_sheet(_safe_sheet_title("Escenarios", used_titles))
    _write_escenarios_sheet(
        scenarios_sheet,
        title=f"Escenarios - {pronostico.nombre}",
        subtitle=subtitle,
        categorias=resultados.get("por_categoria") or [],
    )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"pronostico_{pronostico.id}_{pronostico.fecha_inicio}_{pronostico.fecha_fin}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _pronosticos_for_user(user):
    return PronosticoGuardado.objects.filter(creado_por=user).select_related("creado_por").prefetch_related("sucursales")


def _get_pronostico_for_user(user, pk: int) -> PronosticoGuardado:
    return get_object_or_404(_pronosticos_for_user(user), pk=pk)


def _build_pronostico_detail_context(pronostico: PronosticoGuardado) -> dict:
    resultados = pronostico.resultado_json or {}
    tiene_real, comparativa, resumen_comparativa = _build_comparativa_real(pronostico, resultados)
    return {
        "pronostico": pronostico,
        "tiene_real": tiene_real,
        "comparativa": comparativa,
        "resumen_comparativa": resumen_comparativa,
        **_empty_result_context(resultados),
    }


def _calcular_y_guardar_sync(*, nombre, fecha_inicio, fecha_fin, sucursal_ids, usuario, skus_incluidos=None):
    resultado = calcular_pronostico(fecha_inicio, fecha_fin, set(sucursal_ids), skus_incluidos=skus_incluidos or None)
    resumen = resultado.get("resumen") or {}
    pronostico = PronosticoGuardado.objects.create(
        nombre=nombre,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        resultado_json=_json_ready(resultado),
        total_piezas=resumen.get("total_piezas") or 0,
        total_ingreso=resumen.get("total_ingreso") or 0,
        creado_por=usuario,
    )
    if sucursal_ids:
        pronostico.sucursales.set(Sucursal.objects.filter(id__in=sucursal_ids))
    return pronostico


def _warn_stale_sales_forecast(request) -> None:
    freshness = queue_forecast_sales_refresh_if_needed(triggered_by_id=request.user.id)
    if freshness.is_fresh:
        return

    messages.warning(
        request,
        (
            "Ventas Point atrasadas: ultimo dia cargado "
            f"{freshness.latest_label}; requerido {freshness.target_label}. "
            f"Se lanzo actualizacion de ventas ({freshness.refresh_days} dias) en paralelo — "
            "el pronostico se calcula con los datos actuales."
        ),
    )


@login_required
def PronosticoVentasView(request):
    if not _can_view_pronostico(request.user):
        raise PermissionDenied("No tienes permisos para ver pronosticos de ventas.")

    branches = Sucursal.objects.filter(activa=True).order_by("nombre")
    source = "POST" if request.method == "POST" else "GET"
    data = request.POST if request.method == "POST" else request.GET
    fecha_inicio_raw = (data.get("fecha_inicio") or "").strip()
    fecha_fin_raw = (data.get("fecha_fin") or "").strip()
    selected_branch_ids = _selected_branch_ids(request, source=source)
    categorias_productos = _catalogo_productos_por_categoria()
    available_skus = {product["sku"] for products in categorias_productos.values() for product in products}
    selected_product_skus = _selected_product_skus(request) if request.method == "POST" else sorted(available_skus)
    form_errors = []

    if request.method == "POST":
        fecha_inicio, fecha_fin, form_errors = _validate_dates(fecha_inicio_raw, fecha_fin_raw)
        selected_product_skus = [sku for sku in selected_product_skus if sku in available_skus]
        if not selected_branch_ids:
            form_errors.append("Selecciona al menos una sucursal activa.")
        if not selected_product_skus:
            form_errors.append("Selecciona al menos un producto para incluir en el pronostico.")
        if not form_errors and fecha_inicio and fecha_fin:
            _warn_stale_sales_forecast(request)
            nombre = f"Pronóstico {fecha_inicio.isoformat()} a {fecha_fin.isoformat()}"
            pronostico = _calcular_y_guardar_sync(
                nombre=nombre,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                sucursal_ids=sorted(selected_branch_ids),
                usuario=request.user,
                skus_incluidos=selected_product_skus,
            )
            return redirect("ventas:pronostico_detalle", pk=pronostico.id)
        for error in form_errors:
            messages.error(request, error)

    context = {
        "branches": branches,
        "selected_branch_ids": selected_branch_ids,
        "fecha_inicio": fecha_inicio_raw,
        "fecha_fin": fecha_fin_raw,
        "categorias_productos": categorias_productos,
        "categorias_principales": {category.upper() for category in CATALOG_CATEGORY_ORDER},
        "selected_product_skus": set(selected_product_skus),
        "form_errors": form_errors,
        "pronosticos_guardados": _pronosticos_for_user(request.user)[:10],
        "sales_freshness": get_forecast_sales_freshness(),
        **_empty_result_context({}),
    }
    return render(request, "ventas/pronostico.html", context)


@login_required
def PronosticoEsperandoView(request, task_id: str):
    if not _can_view_pronostico(request.user):
        raise PermissionDenied("No tienes permisos para ver este pronostico.")
    return render(request, "ventas/pronostico_esperando.html", {"task_id": task_id})


@login_required
def PronosticoStatusView(request, task_id: str):
    if not _can_view_pronostico(request.user):
        raise PermissionDenied("No tienes permisos para ver este pronostico.")
    result = AsyncResult(task_id)
    payload = {"status": result.status, "result": None}
    if result.successful():
        payload["result"] = result.result
    elif result.failed():
        payload["error"] = str(result.info)
    return JsonResponse(payload)


@login_required
@require_POST
def PronosticoGuardarView(request):
    if not _can_view_pronostico(request.user):
        raise PermissionDenied("No tienes permisos para guardar pronosticos de ventas.")

    nombre = (request.POST.get("nombre") or "").strip()
    fecha_inicio_raw = (request.POST.get("fecha_inicio") or "").strip()
    fecha_fin_raw = (request.POST.get("fecha_fin") or "").strip()
    fecha_inicio, fecha_fin, errors = _validate_dates(fecha_inicio_raw, fecha_fin_raw)
    selected_branch_ids = _selected_branch_ids(request, source="POST")
    selected_product_skus = _selected_product_skus(request)
    if not selected_branch_ids:
        errors.append("Selecciona al menos una sucursal activa.")
    if errors:
        for error in errors:
            messages.error(request, error)
        return redirect("ventas:pronostico")

    if not nombre:
        nombre = f"Pronóstico {fecha_inicio} a {fecha_fin}"

    _warn_stale_sales_forecast(request)

    pronostico = _calcular_y_guardar_sync(
        nombre=nombre,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        sucursal_ids=sorted(selected_branch_ids),
        usuario=request.user,
        skus_incluidos=selected_product_skus or None,
    )
    return redirect("ventas:pronostico_detalle", pk=pronostico.id)


@login_required
def PronosticoListaView(request):
    if not _can_view_pronostico(request.user):
        raise PermissionDenied("No tienes permisos para ver pronosticos guardados.")
    return render(request, "ventas/pronostico_lista.html", {"pronosticos": _pronosticos_for_user(request.user)})


@login_required
def PronosticoDetalleView(request, pk: int):
    if not _can_view_pronostico(request.user):
        raise PermissionDenied("No tienes permisos para ver este pronostico.")
    pronostico = _get_pronostico_for_user(request.user, pk)
    return render(request, "ventas/pronostico_detalle.html", _build_pronostico_detail_context(pronostico))


@login_required
def PronosticoExcelView(request, pk: int):
    if not _can_view_pronostico(request.user):
        raise PermissionDenied("No tienes permisos para exportar este pronostico.")
    pronostico = _get_pronostico_for_user(request.user, pk)
    return _build_pronostico_excel_response(pronostico)


@login_required
def PronosticoPrintView(request, pk: int):
    if not _can_view_pronostico(request.user):
        raise PermissionDenied("No tienes permisos para imprimir este pronostico.")
    pronostico = _get_pronostico_for_user(request.user, pk)
    return render(request, "ventas/pronostico_print.html", _build_pronostico_detail_context(pronostico))


@login_required
@require_POST
def PronosticoEliminarView(request, pk: int):
    if not _can_view_pronostico(request.user):
        raise PermissionDenied("No tienes permisos para eliminar este pronostico.")
    pronostico = _get_pronostico_for_user(request.user, pk)
    pronostico.delete()
    messages.success(request, "Pronostico eliminado.")
    return redirect("ventas:pronostico_guardados")
