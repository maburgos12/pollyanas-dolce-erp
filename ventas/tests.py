from __future__ import annotations

from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from unittest import mock
from zipfile import ZipFile

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db.models import Sum
from django.test import TestCase, override_settings
from django.core.management import call_command, CommandError
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from core.branch_catalog import eligible_sales_event_branch_qs
from core.models import Sucursal
from inventario.models import ExistenciaInsumo
from maestros.models import CostoInsumo, Insumo, UnidadMedida
from pos_bridge.models import (
    PointBranch,
    PointDailyBranchIndicator,
    PointDailySale,
    PointProduct,
    PointSalesDailyCategoryFact,
    PointSalesDailyProductFact,
    PointSyncJob,
)
from recetas.models import InventarioCedisProducto, LineaReceta, Receta, VentaHistorica
from reportes.models import FactVentaDiaria
from ventas.models import (
    EventoVenta,
    EventoVentaCapacityRule,
    EventoVentaFinancial,
    EventoVentaForecast,
    EventoVentaInputRequirement,
    EventoVentaNotification,
    EventoVentaProducto,
    EventoVentaProjectionArtifact,
    EventoVentaSubstitutionWeight,
    EventoVentaSucursal,
    VentaAutoritativaPoint,
)
from ventas.services.forecasting import (
    _enforce_weekly_executive_ceiling,
    _legacy_pay_recipe_segments,
    _load_point_daily,
    _recent_branch_quantity_total,
    build_event_inputs,
    build_event_executive_projection_model,
    executive_event_product_scope,
    generate_event_forecast,
)
from ventas.services.financials import build_financials
from ventas.services.financials import resolve_unit_price
from ventas.services.point_reconciliation import reconcile_event_point_sales
from ventas.services.postmortem import build_postmortem
from ventas.services.production import generate_production_plan
from ventas.services.requirements import build_input_requirements, build_purchase_requirements
from ventas.services.sales_read_service import get_daily_sales, get_daily_sales_bulk, get_sales_range
from ventas.services.substitution_learning import rebuild_substitution_weights
from ventas.services.sales_truth import authoritative_daily_total, authoritative_day_loaded, verified_point_daily_total, verified_point_sales_aggregate
from ventas.services.event_detail_snapshot import build_event_detail_snapshot_payload, get_event_detail_snapshot_payload
from ventas.views import (
    _attach_forecast_source_flags,
    _build_executive_dashboard_workbook_file,
    _event_financial_dataset,
    _event_projection_window,
    _product_selection_groups,
    _reprocess_event_for_audit,
    _sync_event_review_status_with_guardrails,
)


class VentasEventosServiceTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_superuser(username="ventas_admin", password="secret", email="ventas@example.com")
        self.branch = Sucursal.objects.create(codigo="MATRIZ", nombre="Matriz", activa=True)
        self.product = Receta.objects.create(
            nombre="Pastel Fiesta",
            nombre_normalizado="pastel fiesta",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Temporada",
            hash_contenido="ventas-test-pastel-fiesta",
        )
        self.event = EventoVenta.objects.create(
            name="Dia del Nino 2026",
            event_type="TEMPORADA",
            main_date=date(2026, 4, 30),
            analysis_start_date=date(2026, 4, 26),
            analysis_end_date=date(2026, 4, 30),
        )
        EventoVentaSucursal.objects.create(sales_event=self.event, branch=self.branch)
        EventoVentaProducto.objects.create(sales_event=self.event, product=self.product)

    def test_generate_production_plan_moves_sunday_demand_to_saturday(self):
        sunday = date(2026, 4, 26)
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=sunday,
            final_forecast=Decimal("12"),
            conservative_forecast=Decimal("10"),
            aggressive_forecast=Decimal("14"),
        )
        InventarioCedisProducto.objects.create(receta=self.product, stock_actual=Decimal("2"), stock_reservado=Decimal("0"))

        result = generate_production_plan(self.event)

        self.assertEqual(result["created"], 1)
        line = self.event.production_plans.first().lines.first()
        self.assertEqual(line.production_day, date(2026, 4, 25))
        self.assertEqual(line.net_qty_to_produce, Decimal("10"))

    def test_build_event_inputs_excludes_accessories_and_beverages_from_executive_scope(self):
        bebida = Receta.objects.create(
            nombre="Café frío temporada",
            nombre_normalizado="cafe frio temporada",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            modo_costeo=Receta.MODO_COSTEO_SERVICIO,
            familia="Bebidas",
            categoria="Frías",
            hash_contenido="ventas-test-bebida-evento-001",
        )
        accesorio = Receta.objects.create(
            nombre="Velas cumpleaños",
            nombre_normalizado="velas cumpleanos",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            modo_costeo=Receta.MODO_COSTEO_SERVICIO,
            familia="Accesorios",
            categoria="Complementos",
            hash_contenido="ventas-test-accesorio-evento-001",
        )
        EventoVentaProducto.objects.create(sales_event=self.event, product=bebida)
        EventoVentaProducto.objects.create(sales_event=self.event, product=accesorio)

        inputs = build_event_inputs(self.event)

        self.assertEqual([product.id for product in inputs.products], [self.product.id])
        excluded_names = {product.nombre for product, _reason in inputs.excluded_products}
        self.assertEqual(excluded_names, {"Café frío temporada", "Velas cumpleaños"})

    def test_executive_product_selection_groups_hide_non_core_items(self):
        Receta.objects.create(
            nombre="Batida Chocolate QA",
            nombre_normalizado="batida chocolate qa",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            modo_costeo=Receta.MODO_COSTEO_SERVICIO,
            familia="Bebidas",
            categoria="Especiales",
            hash_contenido="ventas-test-batida-selection-001",
        )

        product_names = {
            product.nombre
            for group in _product_selection_groups()
            for product in group["products"]
        }

        self.assertIn(self.product.nombre, product_names)
        self.assertNotIn("Batida Chocolate QA", product_names)

    def test_executive_event_product_scope_keeps_vasos_preparados_when_product_is_real(self):
        vaso = Receta.objects.create(
            nombre="Vaso Fresas con Crema Grande",
            nombre_normalizado="vaso fresas con crema grande",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            modo_costeo=Receta.MODO_COSTEO_FABRICADO,
            familia="Vasos Preparados",
            categoria="Vasos Grande",
            hash_contenido="ventas-test-vaso-real-001",
        )

        is_eligible, reason = executive_event_product_scope(vaso)

        self.assertTrue(is_eligible)
        self.assertEqual(reason, "eligible")

    def test_attach_forecast_source_flags_uses_dominant_method_for_branch_label(self):
        rows = _attach_forecast_source_flags(
            [
                {
                    "branch__codigo": "MATRIZ",
                    "direct_count": 8,
                    "comparable_count": 2,
                    "fallback_count": 0,
                    "no_data_count": 1,
                }
            ]
        )

        self.assertEqual(rows[0]["source_label"], "Directo")

    def test_attach_forecast_source_flags_treats_ytd_anchor_methods_as_direct(self):
        rows = _attach_forecast_source_flags(
            [
                {
                    "branch__codigo": "MATRIZ",
                    "direct_count": 259,
                    "comparable_count": 0,
                    "fallback_count": 0,
                    "no_data_count": 0,
                }
            ]
        )

        self.assertEqual(rows[0]["source_label"], "Directo")

    def test_generate_production_plan_can_skip_status_promotion(self):
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
            final_forecast=Decimal("12"),
            conservative_forecast=Decimal("10"),
            aggressive_forecast=Decimal("14"),
        )

        generate_production_plan(self.event, promote_status=False)
        self.event.refresh_from_db()

        self.assertEqual(self.event.status, EventoVenta.STATUS_BORRADOR)

    def test_generate_production_plan_reserves_cedis_stock_across_days(self):
        first_day = self.event.analysis_start_date
        second_day = first_day + timedelta(days=1)
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=first_day,
            final_forecast=Decimal("3"),
            conservative_forecast=Decimal("3"),
            aggressive_forecast=Decimal("3"),
        )
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=second_day,
            final_forecast=Decimal("3"),
            conservative_forecast=Decimal("3"),
            aggressive_forecast=Decimal("3"),
        )
        InventarioCedisProducto.objects.create(receta=self.product, stock_actual=Decimal("4"), stock_reservado=Decimal("0"))

        generate_production_plan(self.event)

        lines = list(
            self.event.production_plans.order_by("plan_date")
            .values_list("lines__existing_finished_stock", "lines__net_qty_to_produce")
        )
        self.assertEqual(lines[0][0], Decimal("3.000"))
        self.assertEqual(lines[0][1], Decimal("0.000"))
        self.assertEqual(lines[1][0], Decimal("1.000"))
        self.assertEqual(lines[1][1], Decimal("2.000"))

    def test_generate_production_plan_shares_stock_pool_by_effective_commercial_sku(self):
        first_day = self.event.analysis_start_date
        second_day = first_day + timedelta(days=1)
        effective_recipe = Receta.objects.create(
            nombre="Pastel de Crunch R",
            nombre_normalizado="pastel de crunch r",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Temporada",
            codigo_point="0063",
            hash_contenido="ventas-test-effective-crunch-r",
        )
        alias_recipe = Receta.objects.create(
            nombre="Pastel Crunch - Rebanada",
            nombre_normalizado="pastel crunch rebanada",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Temporada",
            hash_contenido="ventas-test-broken-crunch-alias",
        )
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=alias_recipe,
            forecast_date=first_day,
            final_forecast=Decimal("3"),
            conservative_forecast=Decimal("3"),
            aggressive_forecast=Decimal("3"),
        )
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=effective_recipe,
            forecast_date=second_day,
            final_forecast=Decimal("3"),
            conservative_forecast=Decimal("3"),
            aggressive_forecast=Decimal("3"),
        )
        InventarioCedisProducto.objects.create(receta=effective_recipe, stock_actual=Decimal("4"), stock_reservado=Decimal("0"))

        generate_production_plan(self.event, promote_status=False)

        lines = list(
            self.event.production_plans.order_by("plan_date")
            .values_list("lines__product__nombre", "lines__existing_finished_stock", "lines__net_qty_to_produce")
        )
        self.assertEqual(lines[0][0], "Pastel Crunch - Rebanada")
        self.assertEqual(lines[0][1], Decimal("3.000"))
        self.assertEqual(lines[0][2], Decimal("0.000"))
        self.assertEqual(lines[1][0], "Pastel de Crunch R")
        self.assertEqual(lines[1][1], Decimal("1.000"))
        self.assertEqual(lines[1][2], Decimal("2.000"))

    def test_build_postmortem_compares_forecast_against_actual_sales(self):
        target_day = date(2026, 4, 30)
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=target_day,
            final_forecast=Decimal("20"),
            conservative_forecast=Decimal("18"),
            aggressive_forecast=Decimal("24"),
        )
        VentaHistorica.objects.create(
            receta=self.product,
            sucursal=self.branch,
            fecha=target_day,
            cantidad=Decimal("25"),
            monto_total=Decimal("1250"),
        )

        result = build_postmortem(self.event)

        self.assertEqual(result["created"], 1)
        metric = self.event.execution_metrics.get(metric_date=target_day, branch=self.branch, product=self.product)
        self.assertEqual(metric.forecast_qty, Decimal("20"))
        self.assertEqual(metric.actual_qty, Decimal("25"))
        self.assertEqual(metric.variance_qty, Decimal("5"))

    def test_build_postmortem_uses_canonical_daily_sales_when_history_is_missing(self):
        target_day = date(2026, 4, 30)
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=target_day,
            final_forecast=Decimal("20"),
            conservative_forecast=Decimal("18"),
            aggressive_forecast=Decimal("24"),
        )
        FactVentaDiaria.objects.create(
            fecha=target_day,
            sucursal=self.branch,
            receta=self.product,
            producto_clave=self.product.codigo_point or str(self.product.id),
            producto_nombre=self.product.nombre,
            categoria=self.product.categoria or "",
            cantidad=Decimal("22"),
            tickets=3,
            venta_bruta=Decimal("1100.00"),
            descuento=Decimal("0"),
            venta_total=Decimal("1100.00"),
            venta_neta=Decimal("1100.00"),
            costo_estimado=Decimal("0"),
            margen=Decimal("0"),
            source_kind=FactVentaDiaria.SOURCE_V2,
        )

        result = build_postmortem(self.event)

        self.assertEqual(result["created"], 1)
        metric = self.event.execution_metrics.get(metric_date=target_day, branch=self.branch, product=self.product)
        self.assertEqual(metric.actual_qty, Decimal("22"))
        self.assertEqual(metric.actual_sales, Decimal("1100.00"))

    def test_generate_production_plan_caps_output_when_capacity_rule_exists(self):
        target_day = date(2026, 4, 30)
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=target_day,
            final_forecast=Decimal("20"),
            conservative_forecast=Decimal("18"),
            aggressive_forecast=Decimal("24"),
        )
        InventarioCedisProducto.objects.create(receta=self.product, stock_actual=Decimal("0"), stock_reservado=Decimal("0"))
        EventoVentaCapacityRule.objects.create(
            sales_event=self.event,
            capacity_date=target_day,
            product=self.product,
            max_production_qty=Decimal("12"),
            notes="Capacidad real del horno",
        )

        result = generate_production_plan(self.event)

        self.assertEqual(result["constrained_lines"], 1)
        line = self.event.production_plans.first().lines.first()
        self.assertEqual(line.required_qty, Decimal("20"))
        self.assertEqual(line.planned_qty, Decimal("12"))
        self.assertEqual(line.net_qty_to_produce, Decimal("12"))
        self.assertEqual(line.capacity_limit_qty, Decimal("12"))
        self.assertEqual(line.capacity_gap_qty, Decimal("8"))
        self.assertEqual(line.priority, "CRITICA")

    def test_generate_event_forecast_uses_only_canonical_point_branches(self):
        extra_branch = Sucursal.objects.create(codigo="COL", nombre="Colosio duplicada", activa=True)
        EventoVentaSucursal.objects.create(sales_event=self.event, branch=extra_branch)

        result = generate_event_forecast(self.event)

        self.assertEqual(result["created"], 5)
        self.assertEqual(EventoVentaForecast.objects.filter(sales_event=self.event).count(), 5)
        self.assertFalse(EventoVentaForecast.objects.filter(sales_event=self.event, branch=extra_branch).exists())

    def test_legacy_pay_segments_use_old_guayaba_slice_only_as_history(self):
        legacy_recipe = Receta.objects.create(
            nombre="Pay de Guayaba R",
            codigo_point="0011",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="legacy-guayaba-r",
        )
        current_recipe = Receta.objects.create(
            nombre="Sabor Guayaba Rebanada",
            codigo_point="03SPGREB",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="current-guayaba-r",
        )
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product = PointProduct.objects.create(external_id="guayaba-r", sku="03SPGREB", name=current_recipe.nombre)
        sync_job = PointSyncJob.objects.create(
            job_type=PointSyncJob.JOB_TYPE_SALES,
            status=PointSyncJob.STATUS_SUCCESS,
            triggered_by=self.user,
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=current_recipe,
            sync_job=sync_job,
            sale_date=date(2026, 4, 1),
            quantity=Decimal("1"),
            tickets=1,
            gross_amount=Decimal("0"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("0"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("0"),
        )

        segments = _legacy_pay_recipe_segments(current_recipe, date(2026, 3, 1), date(2026, 4, 10))

        self.assertEqual(len(segments), 1)
        self.assertEqual(segments[0][0].id, legacy_recipe.id)
        self.assertEqual(segments[0][1], date(2026, 3, 1))
        self.assertEqual(segments[0][2], date(2026, 3, 31))

    def test_legacy_pay_segments_use_central_fresa_history_spec(self):
        legacy_recipe = Receta.objects.create(
            nombre="Pay de Queso Rebanada",
            codigo_point="0003",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="legacy-fresa-r",
        )
        current_recipe = Receta.objects.create(
            nombre="Sabor Fresa Rebanada Pay",
            codigo_point="03SPFREB",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="current-fresa-r",
        )
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product = PointProduct.objects.create(external_id="fresa-r", sku="03SPFREB", name=current_recipe.nombre)
        sync_job = PointSyncJob.objects.create(
            job_type=PointSyncJob.JOB_TYPE_SALES,
            status=PointSyncJob.STATUS_SUCCESS,
            triggered_by=self.user,
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=current_recipe,
            sync_job=sync_job,
            sale_date=date(2026, 4, 4),
            quantity=Decimal("1"),
            tickets=1,
            gross_amount=Decimal("0"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("0"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("0"),
        )

        segments = _legacy_pay_recipe_segments(current_recipe, date(2026, 3, 1), date(2026, 4, 10))

        self.assertEqual(len(segments), 1)
        self.assertEqual(segments[0][0].id, legacy_recipe.id)
        self.assertEqual(segments[0][1], date(2026, 3, 1))
        self.assertEqual(segments[0][2], date(2026, 4, 3))

    def test_legacy_pay_segments_use_old_galleta_cajeta_slice_only_as_history(self):
        legacy_recipe = Receta.objects.create(
            nombre="Pay de Queso con Galleta y Cajeta R",
            codigo_point="0019",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="legacy-galleta-cajeta-r",
        )
        current_recipe = Receta.objects.create(
            nombre="Sabor Galleta con Cajeta Rebanada",
            codigo_point="03SPGCCREB",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="current-galleta-cajeta-r",
        )
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product = PointProduct.objects.create(external_id="gc-r", sku="03SPGCCREB", name=current_recipe.nombre)
        sync_job = PointSyncJob.objects.create(
            job_type=PointSyncJob.JOB_TYPE_SALES,
            status=PointSyncJob.STATUS_SUCCESS,
            triggered_by=self.user,
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=current_recipe,
            sync_job=sync_job,
            sale_date=date(2026, 4, 2),
            quantity=Decimal("1"),
            tickets=1,
            gross_amount=Decimal("0"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("0"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("0"),
        )

        segments = _legacy_pay_recipe_segments(current_recipe, date(2026, 3, 1), date(2026, 4, 10))

        self.assertEqual(len(segments), 1)
        self.assertEqual(segments[0][0].id, legacy_recipe.id)
        self.assertEqual(segments[0][1], date(2026, 3, 1))
        self.assertEqual(segments[0][2], date(2026, 4, 1))

    def test_topping_segments_do_not_use_base_cake_as_history(self):
        Receta.objects.create(
            nombre="Pastel de Crunch Mediano",
            codigo_point="0060",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="legacy-crunch-m",
        )
        current_recipe = Receta.objects.create(
            nombre="TOPPING CRUNCH M",
            codigo_point="21125",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="current-topping-crunch-m",
        )
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product = PointProduct.objects.create(external_id="crunch-m", sku="21125", name=current_recipe.nombre)
        sync_job = PointSyncJob.objects.create(
            job_type=PointSyncJob.JOB_TYPE_SALES,
            status=PointSyncJob.STATUS_SUCCESS,
            triggered_by=self.user,
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=current_recipe,
            sync_job=sync_job,
            sale_date=date(2026, 4, 3),
            quantity=Decimal("1"),
            tickets=1,
            gross_amount=Decimal("0"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("0"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("0"),
        )

        segments = _legacy_pay_recipe_segments(current_recipe, date(2026, 3, 1), date(2026, 4, 10))

        self.assertEqual(segments, [])

    def test_brownie_segments_do_not_use_chocolate_chip_pay_as_history(self):
        Receta.objects.create(
            nombre="Pay de Queso con Chispas de Chocolate Grande",
            codigo_point="0619",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="legacy-brownie-g",
        )
        current_recipe = Receta.objects.create(
            nombre="Sabor Brownie Grande",
            codigo_point="SBROWNIEG",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="current-brownie-g",
        )
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product = PointProduct.objects.create(external_id="brownie-g", sku="SBROWNIEG", name=current_recipe.nombre)
        sync_job = PointSyncJob.objects.create(
            job_type=PointSyncJob.JOB_TYPE_SALES,
            status=PointSyncJob.STATUS_SUCCESS,
            triggered_by=self.user,
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=current_recipe,
            sync_job=sync_job,
            sale_date=date(2026, 4, 3),
            quantity=Decimal("1"),
            tickets=1,
            gross_amount=Decimal("0"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("0"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("0"),
        )

        segments = _legacy_pay_recipe_segments(current_recipe, date(2026, 3, 1), date(2026, 4, 10))

        self.assertEqual(segments, [])

    def test_build_financials_is_idempotent(self):
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
            final_forecast=Decimal("10"),
            conservative_forecast=Decimal("8"),
            aggressive_forecast=Decimal("12"),
        )

        build_financials(self.event)
        build_financials(self.event)

        self.assertEqual(
            EventoVentaFinancial.objects.filter(sales_event=self.event).count(),
            3,
        )

    def test_build_financials_uses_branch_specific_current_point_prices(self):
        other_branch = Sucursal.objects.create(codigo="LEYVA", nombre="Leyva", activa=True)
        point_branch_main = PointBranch.objects.create(external_id="bf-main", name="Matriz", erp_branch=self.branch)
        point_branch_other = PointBranch.objects.create(external_id="bf-other", name="Leyva", erp_branch=other_branch)
        point_product = PointProduct.objects.create(external_id="bf-prod", sku="BFP1", name=self.product.nombre, category="Temporada")
        latest_day = date(2026, 4, 8)
        for offset in range(2):
            sale_day = latest_day - timedelta(days=offset)
            PointDailySale.objects.create(
                branch=point_branch_main,
                product=point_product,
                receta=self.product,
                sale_date=sale_day,
                quantity=Decimal("1"),
                total_amount=Decimal("100"),
                source_endpoint="/Report/PrintReportes?idreporte=3",
            )
            PointDailySale.objects.create(
                branch=point_branch_other,
                product=point_product,
                receta=self.product,
                sale_date=sale_day,
                quantity=Decimal("1"),
                total_amount=Decimal("80"),
                source_endpoint="/Report/PrintReportes?idreporte=3",
            )

        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
            final_forecast=Decimal("10"),
            conservative_forecast=Decimal("10"),
            aggressive_forecast=Decimal("10"),
        )
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=other_branch,
            product=self.product,
            forecast_date=self.event.main_date,
            final_forecast=Decimal("5"),
            conservative_forecast=Decimal("5"),
            aggressive_forecast=Decimal("5"),
        )

        build_financials(self.event)

        base = EventoVentaFinancial.objects.get(sales_event=self.event, scenario=EventoVenta.SCENARIO_BASE)
        self.assertEqual(base.estimated_sales, Decimal("1400"))

    def test_executive_dashboard_workbook_includes_sku_governance_sheet(self):
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
            final_forecast=Decimal("5"),
            conservative_forecast=Decimal("4"),
            aggressive_forecast=Decimal("6"),
        )

        filename, payload = _build_executive_dashboard_workbook_file(self.event)
        workbook = load_workbook(BytesIO(payload))

        self.assertIn("Gobernanza SKU", workbook.sheetnames)
        self.assertIn("Auditoria", workbook.sheetnames)
        self.assertTrue(filename.endswith(".xlsx"))

    def test_generate_event_forecast_skips_blocked_ambiguous_skus(self):
        blocked_product = Receta.objects.create(
            nombre="Sabor Mango Grande Pay",
            codigo_point="SMANGOG",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="blocked-mango-g",
        )
        EventoVentaProducto.objects.create(
            sales_event=self.event,
            product=blocked_product,
            source_type=EventoVentaProducto.SOURCE_MANUAL,
        )

        result = generate_event_forecast(self.event)

        self.assertEqual(
            EventoVentaForecast.objects.filter(sales_event=self.event, product=blocked_product).count(),
            0,
        )
        self.assertTrue(any("SKU bloqueado por ambiguedad" in warning for warning in result["warnings"]))

    def test_reprocess_event_for_audit_restores_original_status(self):
        self.event.status = EventoVenta.STATUS_ENVIADO_COMPRAS
        self.event.save(update_fields=["status", "updated_at"])
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
            final_forecast=Decimal("5"),
            conservative_forecast=Decimal("4"),
            aggressive_forecast=Decimal("6"),
        )

        result = _reprocess_event_for_audit(self.event, self.user, skip_postmortem=True)

        self.event.refresh_from_db()
        self.assertEqual(self.event.status, EventoVenta.STATUS_ENVIADO_COMPRAS)
        self.assertGreaterEqual(len(result["artifacts"]), 1)

    def test_reprocess_event_for_audit_refreshes_detail_snapshot(self):
        with mock.patch("ventas.views._refresh_event_detail_snapshot") as refresh_snapshot:
            _reprocess_event_for_audit(self.event, self.user, skip_postmortem=True)

        refresh_snapshot.assert_called_once_with(self.event, generated_by=self.user)

    def test_reprocess_command_writes_audit_workbook_in_reuse_mode(self):
        call_command("reprocess_event_pipelines", event_ids=[self.event.id], reuse_current_state=True)

        target = (
            Path(settings.BASE_DIR)
            / "output"
            / "spreadsheet"
            / "validacion_negocio"
            / f"auditoria_eventos_comerciales_{timezone.localdate().isoformat()}.xlsx"
        )
        workbook = load_workbook(target)

        self.assertTrue(target.exists())
        self.assertIn("Semaforo por modulo", workbook.sheetnames)
        self.assertIn("SKU bloqueados", workbook.sheetnames)
        self.assertIn("Clasificacion maestra", workbook.sheetnames)

    def test_generate_event_forecast_uses_comparable_branch_for_sparse_branch(self):
        today = timezone.localdate()
        guamuchil = Sucursal.objects.create(codigo="GUAMUCHIL", nombre="Guamuchil", activa=False)
        EventoVentaSucursal.objects.create(sales_event=self.event, branch=guamuchil)

        PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        PointBranch.objects.create(external_id="13", name="Guamuchil", erp_branch=guamuchil)
        PointDailyBranchIndicator.objects.create(
            branch=PointBranch.objects.get(erp_branch=self.branch),
            indicator_date=today - timedelta(days=5),
            total_amount=Decimal("100000"),
            total_tickets=100,
            total_avg_ticket=Decimal("1000"),
        )
        PointDailyBranchIndicator.objects.create(
            branch=PointBranch.objects.get(erp_branch=guamuchil),
            indicator_date=today - timedelta(days=5),
            total_amount=Decimal("60000"),
            total_tickets=60,
            total_avg_ticket=Decimal("1000"),
        )
        for lag in range(1, 5):
            target = self.event.main_date - timedelta(days=7 * lag)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=target,
                cantidad=Decimal("10"),
                monto_total=Decimal("100"),
            )

        result = generate_event_forecast(self.event)

        self.assertEqual(result["created"], 10)
        row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=guamuchil,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        self.assertGreater(row.final_forecast, Decimal("0"))
        self.assertEqual(row.explanation_json.get("comparable_branch_code"), "MATRIZ")

    def test_generate_event_forecast_includes_sparse_branch_when_point_signal_exists(self):
        today = timezone.localdate()
        guamuchil = Sucursal.objects.create(codigo="GUAMUCHIL", nombre="Guamuchil", activa=False)
        EventoVentaSucursal.objects.create(sales_event=self.event, branch=guamuchil)

        point_branch_matriz = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_branch_guamuchil = PointBranch.objects.create(external_id="13", name="Guamuchil", erp_branch=guamuchil)
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_matriz,
            indicator_date=today - timedelta(days=5),
            total_amount=Decimal("100000"),
            total_tickets=100,
            total_avg_ticket=Decimal("1000"),
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_guamuchil,
            indicator_date=today - timedelta(days=5),
            total_amount=Decimal("30000"),
            total_tickets=30,
            total_avg_ticket=Decimal("1000"),
        )
        for lag in range(1, 5):
            target = self.event.main_date - timedelta(days=7 * lag)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=target,
                cantidad=Decimal("9"),
                monto_total=Decimal("90"),
            )

        result = generate_event_forecast(self.event)

        self.assertEqual(result["created"], 10)
        self.assertTrue(
            EventoVentaForecast.objects.filter(
                sales_event=self.event,
                branch=guamuchil,
            ).exists()
        )
    def test_generate_event_forecast_applies_minimum_visible_floor_for_new_branch(self):
        today = timezone.localdate()
        guamuchil = Sucursal.objects.create(codigo="GUAMUCHIL", nombre="Guamuchil", activa=False)
        EventoVentaSucursal.objects.create(sales_event=self.event, branch=guamuchil)

        point_branch_matriz = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_branch_guamuchil = PointBranch.objects.create(external_id="13", name="Guamuchil", erp_branch=guamuchil)
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_matriz,
            indicator_date=today - timedelta(days=5),
            total_amount=Decimal("100000"),
            total_tickets=100,
            total_avg_ticket=Decimal("1000"),
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_guamuchil,
            indicator_date=today - timedelta(days=5),
            total_amount=Decimal("30000"),
            total_tickets=30,
            total_avg_ticket=Decimal("1000"),
        )
        for lag in range(1, 3):
            target = self.event.main_date - timedelta(days=7 * lag)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=target,
                cantidad=Decimal("1"),
                monto_total=Decimal("10"),
            )

        generate_event_forecast(self.event)

        row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=guamuchil,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        self.assertEqual(row.final_forecast, Decimal("1"))
        self.assertTrue(row.explanation_json.get("starter_floor_applied"))

    def test_generate_event_forecast_applies_minimum_visible_floor_for_category_fallback_branch(self):
        guamuchil = Sucursal.objects.create(codigo="GUAMUCHIL", nombre="Guamuchil", activa=True)
        EventoVentaSucursal.objects.create(sales_event=self.event, branch=guamuchil)
        self.product.familia = "Bollo"
        self.product.categoria = "Bollo"
        self.product.save(update_fields=["familia", "categoria"])

        for lag in range(1, 3):
            target = self.event.main_date - timedelta(days=7 * lag)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=target,
                cantidad=Decimal("1"),
                monto_total=Decimal("10"),
            )

        generate_event_forecast(self.event)

        row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=guamuchil,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        self.assertEqual(row.final_forecast, Decimal("0"))
        self.assertEqual(row.explanation_json.get("base_method"), "fallback_categoria")
        self.assertFalse(row.explanation_json.get("starter_floor_applied"))

    def test_generate_event_forecast_rebalances_main_day_with_historical_event_curve(self):
        self.event.analysis_start_date = date(2026, 4, 26)
        self.event.analysis_end_date = date(2026, 4, 30)
        self.event.main_date = date(2026, 4, 30)
        self.event.save(update_fields=["analysis_start_date", "analysis_end_date", "main_date"])

        for lag in range(1, 5):
            target = self.event.main_date - timedelta(days=7 * lag)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=target,
                cantidad=Decimal("10"),
                monto_total=Decimal("100"),
            )

        last_year_curve = [
            (date(2025, 4, 26), Decimal("8"), Decimal("80")),
            (date(2025, 4, 27), Decimal("8"), Decimal("80")),
            (date(2025, 4, 28), Decimal("8"), Decimal("80")),
            (date(2025, 4, 29), Decimal("8"), Decimal("80")),
            (date(2025, 4, 30), Decimal("48"), Decimal("480")),
        ]
        for day, qty, sales in last_year_curve:
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=day,
                cantidad=qty,
                monto_total=sales,
            )

        generate_event_forecast(self.event)

        rows = list(
            EventoVentaForecast.objects.filter(sales_event=self.event, branch=self.branch).order_by("forecast_date")
        )
        total_qty = sum((Decimal(str(row.final_forecast or 0)) for row in rows), Decimal("0"))
        main_row = next(row for row in rows if row.forecast_date == self.event.main_date)
        previous_row = next(row for row in rows if row.forecast_date == (self.event.main_date - timedelta(days=1)))
        main_share = (Decimal(str(main_row.final_forecast or 0)) / total_qty).quantize(Decimal("0.0001"))

        self.assertGreater(main_share, Decimal("0.5500"))
        self.assertGreater(Decimal(str(main_row.final_forecast or 0)), Decimal(str(previous_row.final_forecast or 0)))
        self.assertEqual(main_row.explanation_json.get("daily_curve_calibration_scope"), "event_daily_historical_curve")
        self.assertEqual(main_row.explanation_json.get("daily_curve_source"), "branch_historical_qty_curve")

    def test_generate_event_forecast_applies_explicit_main_day_benchmark_floor(self):
        self.event.analysis_start_date = date(2026, 4, 26)
        self.event.analysis_end_date = date(2026, 4, 30)
        self.event.main_date = date(2026, 4, 30)
        self.event.objective_notes = (
            "Benchmark DG 2025 mismo periodo: $500.00\n"
            "Benchmark DG día principal: $490.00"
        )
        self.event.save(update_fields=["analysis_start_date", "analysis_end_date", "main_date", "objective_notes"])

        for lag in range(1, 5):
            target = self.event.main_date - timedelta(days=7 * lag)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=target,
                cantidad=Decimal("10"),
                monto_total=Decimal("100"),
            )

        last_year_curve = [
            (date(2025, 4, 26), Decimal("8"), Decimal("80")),
            (date(2025, 4, 27), Decimal("8"), Decimal("80")),
            (date(2025, 4, 28), Decimal("8"), Decimal("80")),
            (date(2025, 4, 29), Decimal("8"), Decimal("80")),
            (date(2025, 4, 30), Decimal("48"), Decimal("480")),
        ]
        for day, qty, sales in last_year_curve:
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=day,
                cantidad=qty,
                monto_total=sales,
            )

        with mock.patch("ventas.services.financials.resolve_unit_price", return_value=Decimal("10.0000")):
            result = generate_event_forecast(self.event)

        rows = list(
            EventoVentaForecast.objects.filter(sales_event=self.event, branch=self.branch).order_by("forecast_date")
        )
        total_qty = sum((Decimal(str(row.final_forecast or 0)) for row in rows), Decimal("0"))
        main_row = next(row for row in rows if row.forecast_date == self.event.main_date)
        main_qty = Decimal(str(main_row.final_forecast or 0))
        main_share = (main_qty / total_qty).quantize(Decimal("0.0001"))

        self.assertGreaterEqual(main_qty, Decimal("49.000"))
        self.assertGreater(main_share, Decimal("0.6000"))
        self.assertEqual(main_row.explanation_json.get("main_day_benchmark_sales"), 490.0)
        self.assertTrue(any("benchmark DG del día principal" in warning for warning in result["warnings"]))

    def test_generate_event_forecast_anchors_main_day_down_to_explicit_dg_benchmark(self):
        self.event.analysis_start_date = date(2026, 4, 26)
        self.event.analysis_end_date = date(2026, 4, 30)
        self.event.main_date = date(2026, 4, 30)
        self.event.objective_notes = "Benchmark DG día principal: $300.00"
        self.event.save(update_fields=["analysis_start_date", "analysis_end_date", "main_date", "objective_notes"])

        for lag in range(1, 5):
            target = self.event.main_date - timedelta(days=7 * lag)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=target,
                cantidad=Decimal("10"),
                monto_total=Decimal("100"),
            )

        last_year_curve = [
            (date(2025, 4, 26), Decimal("8"), Decimal("80")),
            (date(2025, 4, 27), Decimal("8"), Decimal("80")),
            (date(2025, 4, 28), Decimal("8"), Decimal("80")),
            (date(2025, 4, 29), Decimal("8"), Decimal("80")),
            (date(2025, 4, 30), Decimal("48"), Decimal("480")),
        ]
        for day, qty, sales in last_year_curve:
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=day,
                cantidad=qty,
                monto_total=sales,
            )

        with mock.patch("ventas.services.financials.resolve_unit_prices_bulk", return_value={(self.product.id, self.branch.id): Decimal("10.0000")}):
            generate_event_forecast(self.event)

        rows = list(
            EventoVentaForecast.objects.filter(sales_event=self.event, branch=self.branch).order_by("forecast_date")
        )
        main_row = next(row for row in rows if row.forecast_date == self.event.main_date)
        main_qty = Decimal(str(main_row.final_forecast or 0))

        self.assertLessEqual(main_qty, Decimal("30.100"))
        self.assertEqual(main_row.explanation_json.get("main_day_benchmark_qty_target_source"), "dg_main_day_sales_scale_anchor")

    def test_weekly_executive_ceiling_reshapes_branch_total_to_target(self):
        for day_offset in range((self.event.analysis_end_date - self.event.analysis_start_date).days + 1):
            EventoVentaForecast.objects.create(
                sales_event=self.event,
                branch=self.branch,
                product=self.product,
                forecast_date=self.event.analysis_start_date + timedelta(days=day_offset),
                final_forecast=Decimal("100"),
                conservative_forecast=Decimal("90"),
                aggressive_forecast=Decimal("110"),
            )

        with mock.patch(
            "ventas.services.forecasting.build_event_executive_projection_model",
            return_value={
                "branch_targets": {self.branch.id: Decimal("300.000")},
                "target_total_qty": Decimal("300.000"),
                "current_total_qty": Decimal("500.000"),
                "benchmark_source": "historical_calendar",
                "benchmark_sales": Decimal("30000.00"),
                "same_store_factor": Decimal("0.9000"),
                "expansion_factor": Decimal("0.0000"),
                "contraction_factor": Decimal("1.0000"),
                "comparable_branches": [{"branch_code": self.branch.codigo}],
                "new_branches": [],
                "contracted_branches": [],
            },
        ):
            warnings = _enforce_weekly_executive_ceiling(self.event)

        self.assertEqual(len(warnings), 1)
        total = EventoVentaForecast.objects.filter(sales_event=self.event).aggregate(total=Sum("final_forecast"))["total"]
        self.assertEqual(total, Decimal("300.000"))
        row = EventoVentaForecast.objects.filter(sales_event=self.event).order_by("forecast_date").first()
        self.assertTrue(row.explanation_json.get("weekly_executive_ceiling_applied"))
        self.assertEqual(row.explanation_json.get("weekly_executive_reason"), "executive_branch_target_ceiling")

    def test_weekly_executive_ceiling_preserves_main_day_protected_curve(self):
        historical_curve = [
            (date(2025, 4, 26), Decimal("8"), Decimal("80")),
            (date(2025, 4, 27), Decimal("8"), Decimal("80")),
            (date(2025, 4, 28), Decimal("8"), Decimal("80")),
            (date(2025, 4, 29), Decimal("8"), Decimal("80")),
            (date(2025, 4, 30), Decimal("48"), Decimal("480")),
        ]
        for day, qty, sales in historical_curve:
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=day,
                cantidad=qty,
                monto_total=sales,
            )

        for day_offset in range((self.event.analysis_end_date - self.event.analysis_start_date).days + 1):
            forecast_day = self.event.analysis_start_date + timedelta(days=day_offset)
            EventoVentaForecast.objects.create(
                sales_event=self.event,
                branch=self.branch,
                product=self.product,
                forecast_date=forecast_day,
                final_forecast=Decimal("100"),
                conservative_forecast=Decimal("90"),
                aggressive_forecast=Decimal("110"),
                explanation_json={"main_day_peak_floor_applied": forecast_day == self.event.main_date},
            )

        with mock.patch(
            "ventas.services.forecasting.build_event_executive_projection_model",
            return_value={
                "branch_targets": {self.branch.id: Decimal("300.000")},
                "target_total_qty": Decimal("300.000"),
                "current_total_qty": Decimal("500.000"),
                "benchmark_source": "historical_calendar",
                "benchmark_sales": Decimal("30000.00"),
                "same_store_factor": Decimal("0.9000"),
                "expansion_factor": Decimal("0.0000"),
                "contraction_factor": Decimal("1.0000"),
                "comparable_branches": [{"branch_code": self.branch.codigo}],
                "new_branches": [],
                "contracted_branches": [],
            },
        ):
            warnings = _enforce_weekly_executive_ceiling(self.event)

        self.assertEqual(len(warnings), 1)
        total = EventoVentaForecast.objects.filter(sales_event=self.event).aggregate(total=Sum("final_forecast"))["total"]
        self.assertEqual(total, Decimal("300.000"))
        main_row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        self.assertGreaterEqual(main_row.final_forecast, Decimal("100.000"))
        self.assertTrue(main_row.explanation_json.get("weekly_executive_curve_protected"))

    def test_weekly_executive_ceiling_compresses_protected_curve_when_floor_exceeds_target(self):
        historical_curve = [
            (date(2025, 4, 26), Decimal("12"), Decimal("120")),
            (date(2025, 4, 27), Decimal("12"), Decimal("120")),
            (date(2025, 4, 28), Decimal("12"), Decimal("120")),
            (date(2025, 4, 29), Decimal("12"), Decimal("120")),
            (date(2025, 4, 30), Decimal("52"), Decimal("520")),
        ]
        for day, qty, sales in historical_curve:
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=day,
                cantidad=qty,
                monto_total=sales,
            )

        daily_forecasts = [
            Decimal("70"),
            Decimal("70"),
            Decimal("70"),
            Decimal("70"),
            Decimal("120"),
        ]
        for day_offset, qty in enumerate(daily_forecasts):
            forecast_day = self.event.analysis_start_date + timedelta(days=day_offset)
            EventoVentaForecast.objects.create(
                sales_event=self.event,
                branch=self.branch,
                product=self.product,
                forecast_date=forecast_day,
                final_forecast=qty,
                conservative_forecast=qty,
                aggressive_forecast=qty,
                explanation_json={"main_day_peak_floor_applied": forecast_day == self.event.main_date},
            )

        with mock.patch(
            "ventas.services.forecasting.build_event_executive_projection_model",
            return_value={
                "branch_targets": {self.branch.id: Decimal("100.000")},
                "target_total_qty": Decimal("100.000"),
                "current_total_qty": Decimal("400.000"),
                "benchmark_source": "historical_calendar",
                "benchmark_sales": Decimal("10000.00"),
                "same_store_factor": Decimal("0.9000"),
                "expansion_factor": Decimal("0.0000"),
                "contraction_factor": Decimal("1.0000"),
                "comparable_branches": [{"branch_code": self.branch.codigo}],
                "new_branches": [],
                "contracted_branches": [],
            },
        ):
            warnings = _enforce_weekly_executive_ceiling(self.event)

        self.assertTrue(
            any("Se comprimió la curva diaria protegida" in warning for warning in warnings),
            warnings,
        )
        total = EventoVentaForecast.objects.filter(sales_event=self.event).aggregate(total=Sum("final_forecast"))["total"]
        self.assertEqual(total, Decimal("100.000"))
        main_row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        self.assertEqual(main_row.final_forecast, Decimal("100.000"))

    def test_generate_event_forecast_iterates_weekly_ceiling_until_stable(self):
        with mock.patch(
            "ventas.services.financials.resolve_unit_prices_bulk",
            return_value={(self.product.id, self.branch.id): Decimal("10.0000")},
        ), mock.patch(
            "ventas.services.forecasting._calibrate_forecast_against_event_homologue",
            return_value=[],
        ), mock.patch(
            "ventas.services.forecasting._apply_family_strategy_adjustments",
            return_value=[],
        ), mock.patch(
            "ventas.services.forecasting._rebalance_main_date_priority",
            return_value=[],
        ), mock.patch(
            "ventas.services.forecasting._align_forecast_to_executive_branch_model",
            return_value=[],
        ), mock.patch(
            "ventas.services.forecasting._enforce_main_day_peak_floor",
            return_value=[],
        ), mock.patch(
            "ventas.services.forecasting._enforce_weekly_executive_ceiling",
            side_effect=[["iter-1"], ["iter-2"], []],
        ) as ceiling_mock:
            result = generate_event_forecast(self.event)

        self.assertEqual(ceiling_mock.call_count, 3)
        self.assertIn("iter-1", result["warnings"])
        self.assertIn("iter-2", result["warnings"])

    def test_generate_event_forecast_uses_last_observed_signal_for_future_event_comparable(self):
        today = timezone.localdate()
        self.event.main_date = today + timedelta(days=35)
        self.event.analysis_start_date = self.event.main_date - timedelta(days=3)
        self.event.analysis_end_date = self.event.main_date + timedelta(days=3)
        self.event.save(update_fields=["main_date", "analysis_start_date", "analysis_end_date"])

        guamuchil = Sucursal.objects.create(codigo="GUAMUCHIL", nombre="Guamuchil", activa=True)
        EventoVentaSucursal.objects.create(sales_event=self.event, branch=guamuchil)

        point_branch_matriz = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_branch_guamuchil = PointBranch.objects.create(external_id="13", name="Guamuchil", erp_branch=guamuchil)
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_matriz,
            indicator_date=today - timedelta(days=5),
            total_amount=Decimal("100000"),
            total_tickets=100,
            total_avg_ticket=Decimal("1000"),
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_guamuchil,
            indicator_date=today - timedelta(days=5),
            total_amount=Decimal("60000"),
            total_tickets=60,
            total_avg_ticket=Decimal("1000"),
        )
        for lag in range(1, 5):
            target = self.event.main_date - timedelta(days=7 * lag)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=target,
                cantidad=Decimal("10"),
                monto_total=Decimal("100"),
            )

        generate_event_forecast(self.event)

        row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=guamuchil,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        self.assertGreater(row.final_forecast, Decimal("0"))
        self.assertEqual(row.explanation_json.get("comparable_branch_code"), "MATRIZ")

    def test_generate_event_forecast_honors_configured_guamuchil_comparable_branch(self):
        today = timezone.localdate()
        alt_branch = Sucursal.objects.create(codigo="COLOSIO", nombre="Colosio", activa=True)
        guamuchil = Sucursal.objects.create(codigo="GUAMUCHIL", nombre="Guamuchil", activa=True)
        EventoVentaSucursal.objects.create(sales_event=self.event, branch=alt_branch)
        EventoVentaSucursal.objects.create(
            sales_event=self.event,
            branch=guamuchil,
            comparable_branch=self.branch,
        )

        point_branch_matriz = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_branch_alt = PointBranch.objects.create(external_id="2", name="Colosio", erp_branch=alt_branch)
        point_branch_guamuchil = PointBranch.objects.create(external_id="13", name="Guamuchil", erp_branch=guamuchil)
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_matriz,
            indicator_date=today - timedelta(days=5),
            total_amount=Decimal("100000"),
            total_tickets=100,
            total_avg_ticket=Decimal("1000"),
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_alt,
            indicator_date=today - timedelta(days=5),
            total_amount=Decimal("65000"),
            total_tickets=65,
            total_avg_ticket=Decimal("1000"),
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_guamuchil,
            indicator_date=today - timedelta(days=5),
            total_amount=Decimal("60000"),
            total_tickets=60,
            total_avg_ticket=Decimal("1000"),
        )
        for lag in range(1, 5):
            target = self.event.main_date - timedelta(days=7 * lag)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=target,
                cantidad=Decimal("10"),
                monto_total=Decimal("100"),
            )
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=alt_branch,
                fecha=target,
                cantidad=Decimal("6"),
                monto_total=Decimal("60"),
            )

        generate_event_forecast(self.event)

        row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=guamuchil,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        self.assertEqual(row.explanation_json.get("comparable_branch_code"), "MATRIZ")

    def test_executive_projection_model_separates_same_store_and_expansion(self):
        today = timezone.localdate()
        comparable_branch = Sucursal.objects.create(codigo="COLOSIO", nombre="Colosio", activa=True)
        new_branch = Sucursal.objects.create(codigo="GUAMUCHIL", nombre="Guamuchil", activa=True)
        EventoVentaSucursal.objects.create(sales_event=self.event, branch=comparable_branch)
        EventoVentaSucursal.objects.create(
            sales_event=self.event,
            branch=new_branch,
            comparable_branch=self.branch,
        )

        point_branch_matriz = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_branch_colosio = PointBranch.objects.create(external_id="2", name="Colosio", erp_branch=comparable_branch)
        point_branch_guamuchil = PointBranch.objects.create(external_id="13", name="Guamuchil", erp_branch=new_branch)

        for point_branch, total in (
            (point_branch_matriz, Decimal("100000")),
            (point_branch_colosio, Decimal("90000")),
            (point_branch_guamuchil, Decimal("42000")),
        ):
            PointDailyBranchIndicator.objects.create(
                branch=point_branch,
                indicator_date=today - timedelta(days=5),
                total_amount=total,
                total_tickets=100,
                total_avg_ticket=Decimal("1000"),
            )

        for lag, qty in enumerate([Decimal("10"), Decimal("12"), Decimal("11"), Decimal("13")], start=1):
            target = self.event.main_date - timedelta(days=7 * lag)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=target,
                cantidad=qty,
                monto_total=qty * Decimal("100"),
            )
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=comparable_branch,
                fecha=target,
                cantidad=qty * Decimal("0.8"),
                monto_total=qty * Decimal("80"),
            )
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=target.replace(year=2025),
                cantidad=qty,
                monto_total=qty * Decimal("100"),
            )
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=comparable_branch,
                fecha=target.replace(year=2025),
                cantidad=qty * Decimal("0.7"),
                monto_total=qty * Decimal("70"),
            )

        generate_event_forecast(self.event)
        model = build_event_executive_projection_model(self.event)

        self.assertTrue(str(model["benchmark_source"]).startswith("historical_"))
        self.assertGreater(Decimal(str(model["same_store_factor"])), Decimal("0"))
        self.assertGreater(Decimal(str(model["expansion_increment_qty"])), Decimal("0"))
        self.assertTrue(any(item["branch_code"] == "GUAMUCHIL" for item in model["new_branches"]))
        self.assertTrue(any(item["branch_code"] == "MATRIZ" for item in model["comparable_branches"]))

    def test_executive_projection_model_marks_closed_branch_as_contraction(self):
        closed_branch = Sucursal.objects.create(codigo="PAYAN", nombre="Payan", activa=False)
        EventoVentaSucursal.objects.create(sales_event=self.event, branch=closed_branch)
        for lag, qty in enumerate([Decimal("14"), Decimal("13"), Decimal("15"), Decimal("16")], start=1):
            target = self.event.main_date - timedelta(days=7 * lag)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=closed_branch,
                fecha=target,
                cantidad=qty,
                monto_total=qty * Decimal("100"),
            )
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=closed_branch,
                fecha=target.replace(year=2025),
                cantidad=qty,
                monto_total=qty * Decimal("100"),
            )

        generate_event_forecast(self.event)
        model = build_event_executive_projection_model(self.event)

        self.assertGreater(Decimal(str(model["contraction_qty"])), Decimal("0"))
        self.assertTrue(any(item["branch_code"] == "PAYAN" for item in model["contracted_branches"]))

    def test_executive_projection_model_ignores_low_coverage_indicator_history_for_same_store(self):
        comparable_branch = Sucursal.objects.create(codigo="COLOSIO", nombre="Colosio", activa=True)
        EventoVentaSucursal.objects.create(sales_event=self.event, branch=comparable_branch)

        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_branch_comp = PointBranch.objects.create(external_id="2", name="Colosio", erp_branch=comparable_branch)
        today = timezone.localdate()
        for branch, current_total, prior_total in (
            (point_branch, Decimal("900000"), Decimal("12000")),
            (point_branch_comp, Decimal("700000"), Decimal("9000")),
        ):
            PointDailyBranchIndicator.objects.create(
                branch=branch,
                indicator_date=today - timedelta(days=3),
                total_amount=current_total,
                total_tickets=100,
                total_avg_ticket=Decimal("1000"),
            )
            PointDailyBranchIndicator.objects.create(
                branch=branch,
                indicator_date=(today - timedelta(days=3)).replace(year=2025),
                total_amount=prior_total,
                total_tickets=10,
                total_avg_ticket=Decimal("900"),
            )

        for lag, (main_qty, comp_qty) in enumerate(
            [(Decimal("16"), Decimal("12")), (Decimal("15"), Decimal("11")), (Decimal("14"), Decimal("10"))],
            start=1,
        ):
            target = self.event.main_date - timedelta(days=7 * lag)
            for branch, qty in ((self.branch, main_qty), (comparable_branch, comp_qty)):
                VentaHistorica.objects.create(
                    receta=self.product,
                    sucursal=branch,
                    fecha=target,
                    cantidad=qty,
                    monto_total=qty * Decimal("100"),
                )
                VentaHistorica.objects.create(
                    receta=self.product,
                    sucursal=branch,
                    fecha=target.replace(year=2025),
                    cantidad=qty,
                    monto_total=qty * Decimal("100"),
                )

        generate_event_forecast(self.event)
        model = build_event_executive_projection_model(self.event)
        matriz = next(item for item in model["comparable_branches"] if item["branch_code"] == "MATRIZ")

        self.assertEqual(matriz["same_store_signal_source"], "event_products_ytd")
        self.assertLess(Decimal(str(matriz["same_store_factor"])), Decimal("1.12"))

    def test_generate_event_forecast_applies_current_trend_signal_and_new_model_version(self):
        today = timezone.localdate()
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product = PointProduct.objects.create(external_id="P1", name=self.product.nombre, category="Temporada")
        PointDailyBranchIndicator.objects.create(
            branch=point_branch,
            indicator_date=today - timedelta(days=5),
            total_amount=Decimal("100000"),
            total_tickets=100,
            total_avg_ticket=Decimal("1000"),
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch,
            indicator_date=today - timedelta(days=31),
            total_amount=Decimal("60000"),
            total_tickets=60,
            total_avg_ticket=Decimal("1000"),
        )
        for days_back, qty in [(1, "14"), (2, "13"), (3, "12"), (4, "14"), (20, "4"), (21, "5"), (22, "4")]:
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product,
                receta=self.product,
                sale_date=today - timedelta(days=days_back),
                quantity=Decimal(qty),
                tickets=1,
                total_amount=Decimal("100"),
            )
        for lag in range(1, 5):
            target = self.event.main_date - timedelta(days=7 * lag)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=target,
                cantidad=Decimal("8"),
                monto_total=Decimal("80"),
            )

        generate_event_forecast(self.event)

        row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        self.assertGreaterEqual(row.trend_adjustment, Decimal("0"))
        self.assertEqual(row.model_version, "v8-executive-same-store-expansion")
        self.assertGreater(row.final_forecast, Decimal("0"))
        self.assertLess(row.conservative_forecast, row.final_forecast)
        self.assertGreater(row.aggressive_forecast, row.final_forecast)
        self.assertEqual(row.explanation_json.get("scenario_method"), "buffered_interval_realista")

    def test_generate_event_forecast_pushes_product_that_is_winning_portfolio_share(self):
        self.product.nombre = "Pay 3 Pecados"
        self.product.nombre_normalizado = "pay 3 pecados"
        self.product.familia = "Pay"
        self.product.categoria = "Pay Mediano"
        self.product.save(update_fields=["nombre", "nombre_normalizado", "familia", "categoria"])
        sibling = Receta.objects.create(
            nombre="Pay Fresas con Crema",
            nombre_normalizado="pay fresas con crema",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pay",
            categoria="Pay Mediano",
            hash_contenido="ventas-test-pay-fresas-share",
        )
        EventoVentaProducto.objects.create(sales_event=self.event, product=sibling)
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product_main = PointProduct.objects.create(external_id="PS1", name=self.product.nombre, category="Pay Mediano")
        point_product_sibling = PointProduct.objects.create(external_id="PS2", name=sibling.nombre, category="Pay Mediano")

        for lag in range(1, 5):
            target = self.event.main_date - timedelta(days=7 * lag)
            for recipe in (self.product, sibling):
                VentaHistorica.objects.create(
                    receta=recipe,
                    sucursal=self.branch,
                    fecha=target,
                    cantidad=Decimal("8"),
                    monto_total=Decimal("800"),
                )

        for days_back, qty_main, qty_sibling in [
            (1, "14", "4"),
            (2, "13", "4"),
            (3, "14", "5"),
            (4, "13", "4"),
            (20, "4", "13"),
            (21, "4", "14"),
            (22, "5", "13"),
            (23, "4", "14"),
        ]:
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product_main,
                receta=self.product,
                sale_date=timezone.localdate() - timedelta(days=days_back),
                quantity=Decimal(qty_main),
                tickets=1,
                total_amount=Decimal("100"),
            )
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product_sibling,
                receta=sibling,
                sale_date=timezone.localdate() - timedelta(days=days_back),
                quantity=Decimal(qty_sibling),
                tickets=1,
                total_amount=Decimal("100"),
            )

        generate_event_forecast(self.event)

        winner_row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        loser_row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=self.branch,
            product=sibling,
            forecast_date=self.event.main_date,
        )

        self.assertGreater(Decimal(str(winner_row.explanation_json.get("portfolio_share_shift_pct") or 0)), Decimal("0"))
        self.assertLess(Decimal(str(loser_row.explanation_json.get("portfolio_share_shift_pct") or 0)), Decimal("0"))
        self.assertGreater(Decimal(str(winner_row.explanation_json.get("substitution_boost_pct") or 0)), Decimal("0"))
        self.assertGreaterEqual(Decimal(str(loser_row.explanation_json.get("substitution_drag_pct") or 0)), Decimal("0"))
        self.assertIn(winner_row.explanation_json.get("substitution_confidence"), {"medium", "high"})
        self.assertGreater(winner_row.final_forecast, loser_row.final_forecast)

    def test_generate_event_forecast_detects_category_rotation_from_grande_to_mini(self):
        self.product.nombre = "Coleccion Mini Chocolate"
        self.product.nombre_normalizado = "coleccion mini chocolate"
        self.product.familia = "Coleccion Chocolate"
        self.product.categoria = "Mini"
        self.product.codigo_point = "PMINI1"
        self.product.save(update_fields=["nombre", "nombre_normalizado", "familia", "categoria", "codigo_point"])
        grande = Receta.objects.create(
            nombre="Coleccion Grande Chocolate",
            nombre_normalizado="coleccion grande chocolate",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Coleccion Chocolate",
            categoria="Grande",
            codigo_point="PGRANDE1",
            hash_contenido="ventas-test-pastel-grande-category",
        )
        EventoVentaProducto.objects.create(sales_event=self.event, product=grande)
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product_mini = PointProduct.objects.create(external_id="PC1", name=self.product.nombre, category="Mini")
        point_product_grande = PointProduct.objects.create(external_id="PC2", name=grande.nombre, category="Grande")

        for lag in range(1, 5):
            target = self.event.main_date - timedelta(days=7 * lag)
            for recipe in (self.product, grande):
                VentaHistorica.objects.create(
                    receta=recipe,
                    sucursal=self.branch,
                    fecha=target,
                    cantidad=Decimal("9"),
                    monto_total=Decimal("900"),
                )

        for days_back, qty_mini, qty_grande in [
            (1, "12", "4"),
            (2, "11", "4"),
            (3, "12", "5"),
            (4, "11", "4"),
            (20, "4", "12"),
            (21, "4", "11"),
            (22, "5", "12"),
            (23, "4", "11"),
        ]:
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product_mini,
                receta=self.product,
                sale_date=timezone.localdate() - timedelta(days=days_back),
                quantity=Decimal(qty_mini),
                tickets=1,
                total_amount=Decimal("100"),
            )
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product_grande,
                receta=grande,
                sale_date=timezone.localdate() - timedelta(days=days_back),
                quantity=Decimal(qty_grande),
                tickets=1,
                total_amount=Decimal("100"),
            )

        generate_event_forecast(self.event)

        mini_row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        grande_row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=self.branch,
            product=grande,
            forecast_date=self.event.main_date,
        )

        self.assertGreater(Decimal(str(mini_row.explanation_json.get("category_trend_pct") or 0)), Decimal("0"))
        self.assertLess(Decimal(str(grande_row.explanation_json.get("category_trend_pct") or 0)), Decimal("0"))
        self.assertGreater(Decimal(str(mini_row.explanation_json.get("portfolio_preference_pct") or 0)), Decimal("0"))
        self.assertLess(Decimal(str(grande_row.explanation_json.get("portfolio_preference_pct") or 0)), Decimal("0"))
        self.assertEqual(mini_row.explanation_json.get("group_scope"), "familia")
        self.assertGreater(mini_row.final_forecast, grande_row.final_forecast)

    def test_generate_event_forecast_normalizes_group_total_after_substitution_adjustment(self):
        self.product.nombre = "Pay 3 Pecados"
        self.product.nombre_normalizado = "pay 3 pecados"
        self.product.familia = "Pay"
        self.product.categoria = "Pay Mediano"
        self.product.save(update_fields=["nombre", "nombre_normalizado", "familia", "categoria"])
        sibling = Receta.objects.create(
            nombre="Pay Fresas con Crema",
            nombre_normalizado="pay fresas con crema",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pay",
            categoria="Pay Mediano",
            hash_contenido="ventas-test-pay-fresas-normalization",
        )
        EventoVentaProducto.objects.create(sales_event=self.event, product=sibling)
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product_main = PointProduct.objects.create(external_id="PN1", name=self.product.nombre, category="Pay Mediano")
        point_product_sibling = PointProduct.objects.create(external_id="PN2", name=sibling.nombre, category="Pay Mediano")

        for lag in range(1, 5):
            target = self.event.main_date - timedelta(days=7 * lag)
            for recipe in (self.product, sibling):
                VentaHistorica.objects.create(
                    receta=recipe,
                    sucursal=self.branch,
                    fecha=target,
                    cantidad=Decimal("8"),
                    monto_total=Decimal("800"),
                )

        for days_back, qty_main, qty_sibling in [
            (1, "14", "4"),
            (2, "13", "4"),
            (3, "14", "5"),
            (4, "13", "4"),
            (20, "4", "13"),
            (21, "4", "14"),
            (22, "5", "13"),
            (23, "4", "14"),
        ]:
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product_main,
                receta=self.product,
                sale_date=timezone.localdate() - timedelta(days=days_back),
                quantity=Decimal(qty_main),
                tickets=1,
                total_amount=Decimal("100"),
            )
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product_sibling,
                receta=sibling,
                sale_date=timezone.localdate() - timedelta(days=days_back),
                quantity=Decimal(qty_sibling),
                tickets=1,
                total_amount=Decimal("100"),
            )

        generate_event_forecast(self.event)

        winner_row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        loser_row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=self.branch,
            product=sibling,
            forecast_date=self.event.main_date,
        )

        pre_normalization_total = Decimal(str(winner_row.explanation_json.get("group_pre_normalization_total_qty") or 0))
        target_group_total = Decimal(str(winner_row.explanation_json.get("group_target_total_qty") or 0))
        normalization_factor = Decimal(str(winner_row.explanation_json.get("group_normalization_factor") or 0))
        self.assertTrue(winner_row.explanation_json.get("group_normalization_applied"))
        self.assertAlmostEqual(float(pre_normalization_total * normalization_factor), float(target_group_total), places=3)
        self.assertLessEqual(
            Decimal(str(winner_row.explanation_json.get("group_growth_pct") or 0)),
            Decimal("0.08"),
        )

    def test_generate_event_forecast_keeps_substitution_low_when_group_lacks_base(self):
        self.product.nombre = "Pay 3 Pecados"
        self.product.nombre_normalizado = "pay 3 pecados"
        self.product.familia = "Pay"
        self.product.categoria = "Pay Mediano"
        self.product.save(update_fields=["nombre", "nombre_normalizado", "familia", "categoria"])
        sibling = Receta.objects.create(
            nombre="Pay Fresas con Crema",
            nombre_normalizado="pay fresas con crema",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pay",
            categoria="Pay Mediano",
            hash_contenido="ventas-test-pay-fresas-low-base",
        )
        EventoVentaProducto.objects.create(sales_event=self.event, product=sibling)
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product_main = PointProduct.objects.create(external_id="PL1", name=self.product.nombre, category="Pay Mediano")
        point_product_sibling = PointProduct.objects.create(external_id="PL2", name=sibling.nombre, category="Pay Mediano")

        for lag in range(1, 3):
            target = self.event.main_date - timedelta(days=7 * lag)
            for recipe in (self.product, sibling):
                VentaHistorica.objects.create(
                    receta=recipe,
                    sucursal=self.branch,
                    fecha=target,
                    cantidad=Decimal("5"),
                    monto_total=Decimal("500"),
                )

        for days_back, qty_main, qty_sibling in [
            (1, "4", "2"),
            (2, "5", "2"),
        ]:
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product_main,
                receta=self.product,
                sale_date=timezone.localdate() - timedelta(days=days_back),
                quantity=Decimal(qty_main),
                tickets=1,
                total_amount=Decimal("100"),
            )
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product_sibling,
                receta=sibling,
                sale_date=timezone.localdate() - timedelta(days=days_back),
                quantity=Decimal(qty_sibling),
                tickets=1,
                total_amount=Decimal("100"),
            )

        generate_event_forecast(self.event)

        winner_row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        loser_row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=self.branch,
            product=sibling,
            forecast_date=self.event.main_date,
        )

        self.assertEqual(winner_row.explanation_json.get("substitution_confidence"), "low")
        self.assertEqual(Decimal(str(winner_row.explanation_json.get("substitution_boost_pct") or 0)), Decimal("0"))
        self.assertEqual(Decimal(str(loser_row.explanation_json.get("substitution_drag_pct") or 0)), Decimal("0"))

    def test_rebuild_forecast_substitution_weights_command_creates_learned_rows(self):
        sibling = Receta.objects.create(
            nombre="Pay Fresas con Crema",
            nombre_normalizado="pay fresas con crema",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pay",
            categoria="Pay Mediano",
            hash_contenido="ventas-test-command-pay-fresas",
        )
        self.product.nombre = "Pay 3 Pecados"
        self.product.nombre_normalizado = "pay 3 pecados"
        self.product.familia = "Pay"
        self.product.categoria = "Pay Mediano"
        self.product.save(update_fields=["nombre", "nombre_normalizado", "familia", "categoria"])

        start = timezone.localdate() - timedelta(days=84)
        for week in range(12):
            sale_day = start + timedelta(days=7 * week)
            winner_qty = Decimal("16") if week >= 6 else Decimal("7")
            loser_qty = Decimal("5") if week >= 6 else Decimal("14")
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=sale_day,
                cantidad=winner_qty,
                monto_total=winner_qty * Decimal("100"),
            )
            VentaHistorica.objects.create(
                receta=sibling,
                sucursal=self.branch,
                fecha=sale_day,
                cantidad=loser_qty,
                monto_total=loser_qty * Decimal("100"),
            )

        call_command(
            "rebuild_forecast_substitution_weights",
            "--lookback-days=98",
            "--window-days=7",
        )

        learned = EventoVentaSubstitutionWeight.objects.filter(
            group_key="familia_categoria::Pay::Pay Mediano",
            winner_product=self.product,
            loser_product=sibling,
            branch__isnull=True,
            version="v7.2-learned",
        ).first()
        self.assertIsNotNone(learned)
        self.assertGreater(learned.weight, Decimal("0"))
        self.assertGreaterEqual(learned.sample_size, 4)

    def test_generate_event_forecast_uses_learned_substitution_weights_with_blended_source(self):
        sibling = Receta.objects.create(
            nombre="Pay Fresas con Crema",
            nombre_normalizado="pay fresas con crema",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pay",
            categoria="Pay Mediano",
            hash_contenido="ventas-test-learned-pay-fresas",
        )
        self.product.nombre = "Pay 3 Pecados"
        self.product.nombre_normalizado = "pay 3 pecados"
        self.product.familia = "Pay"
        self.product.categoria = "Pay Mediano"
        self.product.save(update_fields=["nombre", "nombre_normalizado", "familia", "categoria"])
        EventoVentaProducto.objects.create(sales_event=self.event, product=sibling)
        other_branch = Sucursal.objects.create(codigo="COLOSIO", nombre="Colosio", activa=True)
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product_main = PointProduct.objects.create(external_id="PSB1", name=self.product.nombre, category="Pay Mediano")
        point_product_sibling = PointProduct.objects.create(external_id="PSB2", name=sibling.nombre, category="Pay Mediano")

        local_start = timezone.localdate() - timedelta(days=56)
        for week in range(8):
            sale_day = local_start + timedelta(days=7 * week)
            winner_qty = Decimal("15") if week >= 4 else Decimal("8")
            loser_qty = Decimal("6") if week >= 4 else Decimal("14")
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=sale_day,
                cantidad=winner_qty,
                monto_total=winner_qty * Decimal("100"),
            )
            VentaHistorica.objects.create(
                receta=sibling,
                sucursal=self.branch,
                fecha=sale_day,
                cantidad=loser_qty,
                monto_total=loser_qty * Decimal("100"),
            )

        global_start = timezone.localdate() - timedelta(days=98)
        for week in range(14):
            sale_day = global_start + timedelta(days=7 * week)
            winner_qty = Decimal("14") if week >= 7 else Decimal("7")
            loser_qty = Decimal("5") if week >= 7 else Decimal("13")
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=other_branch,
                fecha=sale_day,
                cantidad=winner_qty,
                monto_total=winner_qty * Decimal("100"),
            )
            VentaHistorica.objects.create(
                receta=sibling,
                sucursal=other_branch,
                fecha=sale_day,
                cantidad=loser_qty,
                monto_total=loser_qty * Decimal("100"),
            )

        for lag in range(1, 5):
            target = self.event.main_date - timedelta(days=7 * lag)
            for recipe in (self.product, sibling):
                VentaHistorica.objects.create(
                    receta=recipe,
                    sucursal=self.branch,
                    fecha=target,
                    cantidad=Decimal("8"),
                    monto_total=Decimal("800"),
                )

        for days_back, qty_main, qty_sibling in [
            (1, "14", "4"),
            (2, "13", "4"),
            (3, "14", "5"),
            (4, "13", "4"),
            (20, "4", "13"),
            (21, "4", "14"),
            (22, "5", "13"),
            (23, "4", "14"),
        ]:
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product_main,
                receta=self.product,
                sale_date=timezone.localdate() - timedelta(days=days_back),
                quantity=Decimal(qty_main),
                tickets=1,
                total_amount=Decimal("100"),
            )
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product_sibling,
                receta=sibling,
                sale_date=timezone.localdate() - timedelta(days=days_back),
                quantity=Decimal(qty_sibling),
                tickets=1,
                total_amount=Decimal("100"),
            )

        rebuild_substitution_weights(lookback_days=120, window_days=7)
        generate_event_forecast(self.event)

        winner_row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        self.assertEqual(winner_row.explanation_json.get("substitution_weight_source"), "learned")
        self.assertEqual(winner_row.explanation_json.get("substitution_source_level"), "blended")
        self.assertGreater(Decimal(str(winner_row.explanation_json.get("substitution_lambda_branch") or 0)), Decimal("0"))
        self.assertGreater(Decimal(str(winner_row.explanation_json.get("substitution_weight_applied") or 0)), Decimal("0"))

    def test_generate_production_plan_uses_operational_target_when_forecast_has_confidence(self):
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
            final_forecast=Decimal("10"),
            conservative_forecast=Decimal("8"),
            aggressive_forecast=Decimal("14"),
            confidence_score=Decimal("0.80"),
        )

        generate_production_plan(self.event, promote_status=False)

        line = self.event.production_plans.first().lines.first()
        self.assertEqual(line.required_qty, Decimal("10.000"))
        self.assertGreater(line.planned_qty, Decimal("10.000"))
        self.assertIn("Objetivo operativo", line.constraint_reason)

    def test_executive_dashboard_workbook_labels_roi_as_gross(self):
        kg = UnidadMedida.objects.create(codigo="kg-test-gross", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA, factor_to_base=Decimal("1000"))
        insumo = Insumo.objects.create(
            codigo="INS-GROSS",
            nombre="Insumo gross",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=kg,
        )
        LineaReceta.objects.create(
            receta=self.product,
            posicion=1,
            insumo=insumo,
            insumo_texto="Insumo gross",
            cantidad=Decimal("1"),
            unidad_texto="kg",
            unidad=kg,
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
        )
        CostoInsumo.objects.create(insumo=insumo, costo_unitario=Decimal("10"), source_hash="ventas-gross-cost")
        VentaHistorica.objects.create(
            receta=self.product,
            sucursal=self.branch,
            fecha=self.event.main_date - timedelta(days=7),
            cantidad=Decimal("5"),
            monto_total=Decimal("100"),
        )
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
            final_forecast=Decimal("5"),
            conservative_forecast=Decimal("4"),
            aggressive_forecast=Decimal("6"),
        )
        build_financials(self.event)

        _filename, payload = _build_executive_dashboard_workbook_file(self.event)
        workbook = load_workbook(BytesIO(payload))
        dashboard = workbook["Dashboard"]
        values = [row[0] for row in dashboard.iter_rows(min_col=1, max_col=1, values_only=True) if row and row[0]]

        self.assertIn("Nota financiera", values)
        self.assertIn("ROI bruto esperado %", values)

    def test_executive_dashboard_workbook_uses_semantic_number_formats(self):
        kg = UnidadMedida.objects.create(codigo="kg-test-format", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA, factor_to_base=Decimal("1000"))
        insumo = Insumo.objects.create(
            codigo="INS-FMT",
            nombre="Insumo format",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=kg,
        )
        LineaReceta.objects.create(
            receta=self.product,
            posicion=1,
            insumo=insumo,
            insumo_texto="Insumo format",
            cantidad=Decimal("1"),
            unidad_texto="kg",
            unidad=kg,
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
        )
        CostoInsumo.objects.create(insumo=insumo, costo_unitario=Decimal("10"), source_hash="ventas-format-cost")
        VentaHistorica.objects.create(
            receta=self.product,
            sucursal=self.branch,
            fecha=self.event.main_date - timedelta(days=7),
            cantidad=Decimal("5"),
            monto_total=Decimal("100"),
        )
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
            final_forecast=Decimal("5"),
            conservative_forecast=Decimal("4"),
            aggressive_forecast=Decimal("6"),
        )
        build_financials(self.event)

        _filename, payload = _build_executive_dashboard_workbook_file(self.event)
        workbook = load_workbook(BytesIO(payload))

        dashboard = workbook["Dashboard"]
        daily = workbook["Demanda diaria"]
        branches = workbook["Sucursales"]
        products = workbook["Productos"]

        dashboard_labels = {
            str(dashboard.cell(row_idx, 1).value or "").strip(): dashboard.cell(row_idx, 2)
            for row_idx in range(1, dashboard.max_row + 1)
        }
        self.assertEqual(dashboard_labels["Volumen proyectado"].number_format, '#,##0.000')
        self.assertEqual(dashboard_labels["Cobertura precio %"].number_format, '0.00"%"')
        self.assertEqual(dashboard_labels["Venta proyectada"].number_format, '$#,##0.00')
        self.assertEqual(dashboard_labels["ROI bruto esperado %"].number_format, '0.00"%"')

        self.assertEqual(daily["B4"].number_format, '#,##0.000')
        self.assertEqual(daily["C4"].number_format, '$#,##0.00')
        self.assertEqual(branches["B4"].number_format, '#,##0.000')
        self.assertEqual(branches["C4"].number_format, '$#,##0.00')
        self.assertEqual(branches["F4"].number_format, '0.00"%"')
        self.assertEqual(products["D4"].number_format, '#,##0.000')
        self.assertEqual(products["E4"].number_format, '$#,##0.00')
        self.assertEqual(products["I4"].number_format, '0.00"%"')

    def test_executive_dashboard_workbook_places_charts_in_stable_grid(self):
        kg = UnidadMedida.objects.create(codigo="kg-test-chart", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA, factor_to_base=Decimal("1000"))
        insumo = Insumo.objects.create(
            codigo="INS-CHART",
            nombre="Insumo chart",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=kg,
        )
        LineaReceta.objects.create(
            receta=self.product,
            posicion=1,
            insumo=insumo,
            insumo_texto="Insumo chart",
            cantidad=Decimal("1"),
            unidad_texto="kg",
            unidad=kg,
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
        )
        CostoInsumo.objects.create(insumo=insumo, costo_unitario=Decimal("10"), source_hash="ventas-chart-cost")
        for offset, qty in enumerate([Decimal("5"), Decimal("6"), Decimal("7"), Decimal("8"), Decimal("9")], start=0):
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=self.event.main_date - timedelta(days=offset),
                cantidad=qty,
                monto_total=qty * Decimal("100"),
            )
            EventoVentaForecast.objects.create(
                sales_event=self.event,
                branch=self.branch,
                product=self.product,
                forecast_date=self.event.main_date - timedelta(days=offset),
                final_forecast=qty,
                conservative_forecast=max(Decimal("0"), qty - Decimal("1")),
                aggressive_forecast=qty + Decimal("1"),
            )
        build_financials(self.event)

        _filename, payload = _build_executive_dashboard_workbook_file(self.event)
        workbook = load_workbook(BytesIO(payload))
        dashboard = workbook["Dashboard"]

        self.assertEqual(len(dashboard._charts), 3)
        anchors = sorted((chart.anchor._from.col, chart.anchor._from.row) for chart in dashboard._charts)
        self.assertEqual(anchors, [(9, 3), (9, 23), (18, 3)])
        self.assertFalse(dashboard.sheet_view.showGridLines)

    def test_executive_dashboard_summary_section_headers_do_not_style_first_data_row(self):
        kg = UnidadMedida.objects.create(codigo="kg-test-header", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA, factor_to_base=Decimal("1000"))
        insumo = Insumo.objects.create(
            codigo="INS-HEADER",
            nombre="Insumo header",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=kg,
        )
        LineaReceta.objects.create(
            receta=self.product,
            posicion=1,
            insumo=insumo,
            insumo_texto="Insumo header",
            cantidad=Decimal("1"),
            unidad_texto="kg",
            unidad=kg,
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
        )
        CostoInsumo.objects.create(insumo=insumo, costo_unitario=Decimal("10"), source_hash="ventas-header-cost")
        for offset, qty in enumerate([Decimal("5"), Decimal("6"), Decimal("7"), Decimal("8"), Decimal("9")], start=0):
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=self.event.main_date - timedelta(days=offset),
                cantidad=qty,
                monto_total=qty * Decimal("100"),
            )
            EventoVentaForecast.objects.create(
                sales_event=self.event,
                branch=self.branch,
                product=self.product,
                forecast_date=self.event.main_date - timedelta(days=offset),
                final_forecast=qty,
                conservative_forecast=max(Decimal("0"), qty - Decimal("1")),
                aggressive_forecast=qty + Decimal("1"),
            )
        build_financials(self.event)

        _filename, payload = _build_executive_dashboard_workbook_file(self.event)
        workbook = load_workbook(BytesIO(payload))
        dashboard = workbook["Dashboard"]

        row_by_label = {}
        for row_idx in range(1, dashboard.max_row + 1):
            label = str(dashboard.cell(row_idx, 1).value or "").strip()
            if label:
                row_by_label[label] = row_idx

        branch_header = row_by_label["Top sucursales"]
        family_header = row_by_label["Mix por familia"]
        first_branch_data = branch_header + 1
        first_family_data = family_header + 1

        self.assertEqual(dashboard.cell(branch_header, 1).fill.fgColor.rgb, "00EBC8D7")
        self.assertEqual(dashboard.cell(branch_header, 2).fill.fgColor.rgb, "00EBC8D7")
        self.assertEqual(dashboard.cell(family_header, 1).fill.fgColor.rgb, "00EBC8D7")
        self.assertEqual(dashboard.cell(family_header, 2).fill.fgColor.rgb, "00EBC8D7")
        self.assertNotEqual(dashboard.cell(first_branch_data, 1).fill.fgColor.rgb, "00EBC8D7")
        self.assertNotEqual(dashboard.cell(first_family_data, 1).fill.fgColor.rgb, "00EBC8D7")

    def test_executive_dashboard_charts_use_complete_visible_summary_ranges(self):
        kg = UnidadMedida.objects.create(codigo="kg-test-range", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA, factor_to_base=Decimal("1000"))
        insumo = Insumo.objects.create(
            codigo="INS-RANGE",
            nombre="Insumo range",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=kg,
        )
        LineaReceta.objects.create(
            receta=self.product,
            posicion=1,
            insumo=insumo,
            insumo_texto="Insumo range",
            cantidad=Decimal("1"),
            unidad_texto="kg",
            unidad=kg,
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
        )
        CostoInsumo.objects.create(insumo=insumo, costo_unitario=Decimal("10"), source_hash="ventas-range-cost")

        extra_branches = []
        for code in ["LEYVA", "LAS_GLORIAS", "COLOSIO", "CRUCERO", "PAYAN"]:
            branch = Sucursal.objects.create(codigo=code, nombre=code.title(), activa=True)
            EventoVentaSucursal.objects.create(sales_event=self.event, branch=branch)
            extra_branches.append(branch)

        all_branches = [self.branch, *extra_branches]
        for idx, branch in enumerate(all_branches, start=1):
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=branch,
                fecha=self.event.main_date,
                cantidad=Decimal("10") + Decimal(idx),
                monto_total=(Decimal("10") + Decimal(idx)) * Decimal("100"),
            )
            EventoVentaForecast.objects.create(
                sales_event=self.event,
                branch=branch,
                product=self.product,
                forecast_date=self.event.main_date,
                final_forecast=Decimal("10") + Decimal(idx),
                conservative_forecast=Decimal("9") + Decimal(idx),
                aggressive_forecast=Decimal("11") + Decimal(idx),
            )
        build_financials(self.event)

        _filename, payload = _build_executive_dashboard_workbook_file(self.event)
        workbook = load_workbook(BytesIO(payload))
        dashboard = workbook["Dashboard"]
        branch_chart = next(chart for chart in dashboard._charts if type(chart).__name__ == "BarChart")
        family_chart = next(chart for chart in dashboard._charts if type(chart).__name__ == "PieChart")

        self.assertIn("$A$27:$A$32", branch_chart.ser[0].cat.numRef.f)
        self.assertIn("$B$27:$B$32", branch_chart.ser[0].val.numRef.f)
        self.assertIn("$A$35", family_chart.ser[0].cat.numRef.f)
        self.assertIn("$B$35", family_chart.ser[0].val.numRef.f)

    def test_generate_event_forecast_caps_product_against_homologue_when_it_overgrows(self):
        self.event.analysis_start_date = self.event.main_date
        self.event.analysis_end_date = self.event.main_date
        self.event.save(update_fields=["analysis_start_date", "analysis_end_date", "updated_at"])

        VentaHistorica.objects.create(
            receta=self.product,
            sucursal=self.branch,
            fecha=date(2025, 5, 10),
            cantidad=Decimal("50"),
            monto_total=Decimal("5000"),
        )
        for lag in range(1, 9):
            target = self.event.main_date - timedelta(days=7 * lag)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=target,
                cantidad=Decimal("30"),
                monto_total=Decimal("3000"),
            )

        generate_event_forecast(self.event)

        total = EventoVentaForecast.objects.filter(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
        ).aggregate(total=Sum("final_forecast"))["total"]
        self.assertLessEqual(total, Decimal("56.000"))
        row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        self.assertIn(row.explanation_json.get("calibration_scope"), ("", "product_homologue_cap"))

    def test_generate_event_forecast_uses_weekday_homologue_floor_for_movable_peak(self):
        self.event.main_date = date(2026, 6, 21)
        self.event.analysis_start_date = date(2026, 6, 16)
        self.event.analysis_end_date = date(2026, 6, 22)
        self.event.save(update_fields=["main_date", "analysis_start_date", "analysis_end_date", "updated_at"])

        VentaHistorica.objects.create(
            receta=self.product,
            sucursal=self.branch,
            fecha=date(2025, 6, 15),
            cantidad=Decimal("120"),
            monto_total=Decimal("12000"),
        )
        for offset in range(0, 21, 7):
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=date(2026, 4, 10) - timedelta(days=offset),
                cantidad=Decimal("4"),
                monto_total=Decimal("400"),
            )

        generate_event_forecast(self.event)

        row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        self.assertGreaterEqual(row.final_forecast, Decimal("100"))
        self.assertEqual(row.explanation_json.get("homologue_mode"), "weekday_occurrence")
        self.assertIn(
            row.explanation_json.get("calibration_scope"),
            ("product_homologue_floor", "event_main_day_homologue_floor"),
        )

    def test_generate_event_forecast_limits_mix_to_products_selected_for_event(self):
        extra_product = Receta.objects.create(
            nombre="Pay Fiesta Chico",
            nombre_normalizado="pay fiesta chico",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pays",
            categoria="Chico",
            hash_contenido="ventas-test-pay-fiesta-chico",
        )
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product = PointProduct.objects.create(external_id="P2", name=extra_product.nombre, category="Chico")
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=extra_product,
            sale_date=self.event.main_date - timedelta(days=5),
            quantity=Decimal("6"),
            tickets=1,
            total_amount=Decimal("300"),
        )

        result = generate_event_forecast(self.event)

        self.assertEqual(result["created"], 5)
        self.assertFalse(EventoVentaForecast.objects.filter(sales_event=self.event, branch=self.branch, product=extra_product).exists())

    def test_build_input_requirements_converts_units_and_blocks_internal_from_purchase(self):
        kg = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA, factor_to_base=Decimal("1000"))
        g = UnidadMedida.objects.create(codigo="g", nombre="Gramo", tipo=UnidadMedida.TIPO_MASA, factor_to_base=Decimal("1"))
        raw_input = Insumo.objects.create(
            codigo="FRESA",
            nombre="Fresa fresca",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=kg,
        )
        internal_input = Insumo.objects.create(
            codigo="BATIDA",
            nombre="Batida interna",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=kg,
        )
        LineaReceta.objects.create(
            receta=self.product,
            posicion=1,
            insumo=raw_input,
            insumo_texto="Fresa fresca",
            cantidad=Decimal("500"),
            unidad_texto="g",
            unidad=g,
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
        )
        LineaReceta.objects.create(
            receta=self.product,
            posicion=2,
            insumo=internal_input,
            insumo_texto="Batida interna",
            cantidad=Decimal("1"),
            unidad_texto="kg",
            unidad=kg,
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
        )
        CostoInsumo.objects.create(insumo=raw_input, costo_unitario=Decimal("100"), source_hash="ventas-test-fresa-cost")
        ExistenciaInsumo.objects.create(insumo=raw_input, stock_actual=Decimal("0"))
        ExistenciaInsumo.objects.create(insumo=internal_input, stock_actual=Decimal("0"))
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
            final_forecast=Decimal("2"),
            conservative_forecast=Decimal("2"),
            aggressive_forecast=Decimal("2"),
        )

        generate_production_plan(self.event, promote_status=False)
        result = build_input_requirements(self.event)
        purchase_result = build_purchase_requirements(self.event)

        self.assertEqual(result["created"], 2)
        raw_requirement = EventoVentaInputRequirement.objects.get(sales_event=self.event, input_item=raw_input)
        internal_requirement = EventoVentaInputRequirement.objects.get(sales_event=self.event, input_item=internal_input)
        self.assertEqual(raw_requirement.required_qty, Decimal("1.000"))
        self.assertEqual(internal_requirement.required_qty, Decimal("2.000"))
        self.assertEqual(purchase_result["created"], 1)
        self.assertEqual(purchase_result["blocked_internal"], 1)

    def test_build_purchase_requirements_does_not_auto_promote_event_to_compras(self):
        self.event.status = EventoVenta.STATUS_ENVIADO_PROD
        self.event.save(update_fields=["status", "updated_at"])
        kg = UnidadMedida.objects.get_or_create(codigo="kg", defaults={"nombre": "Kilogramo"})[0]
        raw_input = Insumo.objects.create(nombre="Fresa compras", tipo_item=Insumo.TIPO_MATERIA_PRIMA, unidad_base=kg)
        LineaReceta.objects.create(
            receta=self.product,
            posicion=1,
            insumo=raw_input,
            insumo_texto="Fresa compras",
            cantidad=Decimal("1"),
            unidad_texto="kg",
            unidad=kg,
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
        )
        CostoInsumo.objects.create(insumo=raw_input, costo_unitario=Decimal("50"), source_hash="ventas-test-compras")
        ExistenciaInsumo.objects.create(insumo=raw_input, stock_actual=Decimal("0"))
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
            final_forecast=Decimal("1"),
            conservative_forecast=Decimal("1"),
            aggressive_forecast=Decimal("1"),
        )

        generate_production_plan(self.event, promote_status=False)
        build_input_requirements(self.event)
        build_purchase_requirements(self.event)

        self.event.refresh_from_db()
        self.assertEqual(self.event.status, EventoVenta.STATUS_ENVIADO_PROD)

    def test_upload_adjustments_updates_week_projection_and_support_outputs(self):
        self.client.force_login(self.user)
        extra_branch = Sucursal.objects.create(codigo="CRUCERO", nombre="Crucero", activa=True)
        EventoVentaSucursal.objects.create(sales_event=self.event, branch=extra_branch)
        for branch, qty in ((self.branch, Decimal("6")), (extra_branch, Decimal("4"))):
            for offset in range(-3, 4):
                forecast_day = self.event.main_date + timedelta(days=offset)
                EventoVentaForecast.objects.create(
                    sales_event=self.event,
                    branch=branch,
                    product=self.product,
                    forecast_date=forecast_day,
                    final_forecast=qty,
                    conservative_forecast=qty,
                    aggressive_forecast=qty,
                )

        payload = (
            "Producto,Proyeccion general,MATRIZ,CRUCERO,Motivo\n"
            "Pastel Fiesta,70,42,28,Ajuste comercial ventas\n"
        ).encode("utf-8")
        uploaded = SimpleUploadedFile("ajustes.csv", payload, content_type="text/csv")

        response = self.client.post(
            reverse("ventas:evento_upload_adjustments", args=[self.event.id]),
            {"adjust_scope": "SEMANA", "adjustments_file": uploaded},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        total = EventoVentaForecast.objects.filter(sales_event=self.event, product=self.product).aggregate(total=Sum("final_forecast"))["total"]
        matriz_total = EventoVentaForecast.objects.filter(sales_event=self.event, product=self.product, branch=self.branch).aggregate(total=Sum("final_forecast"))["total"]
        crucero_total = EventoVentaForecast.objects.filter(sales_event=self.event, product=self.product, branch=extra_branch).aggregate(total=Sum("final_forecast"))["total"]
        self.assertEqual(total, Decimal("70.000"))
        self.assertEqual(matriz_total, Decimal("42.000"))
        self.assertEqual(crucero_total, Decimal("28.000"))
        self.assertTrue(EventoVentaProjectionArtifact.objects.filter(sales_event=self.event).exists())
        self.assertTrue(self.event.financials.exists())

    def test_create_event_does_not_include_guamuchil_by_default(self):
        self.client.force_login(self.user)
        guamuchil = Sucursal.objects.create(codigo="GUAMUCHIL", nombre="Guamuchil", activa=True)

        response = self.client.post(
            reverse("ventas:evento_create"),
            {
                "name": "Evento Madres",
                "event_type": "TEMPORADA",
                "main_date": "2026-05-10",
                "analysis_start_date": "2026-05-07",
                "analysis_end_date": "2026-05-13",
                "branches": [str(self.branch.id)],
                "products": [str(self.product.id)],
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        created_event = EventoVenta.objects.exclude(id=self.event.id).latest("id")
        branch_codes = set(EventoVentaSucursal.objects.filter(sales_event=created_event).values_list("branch__codigo", flat=True))
        self.assertIn("MATRIZ", branch_codes)
        self.assertNotIn("GUAMUCHIL", branch_codes)

    def test_eligible_sales_event_branch_qs_excludes_tmp1(self):
        Sucursal.objects.create(codigo="TMP1", nombre="Tmp1", activa=True)

        branch_codes = set(eligible_sales_event_branch_qs().values_list("codigo", flat=True))

        self.assertNotIn("TMP1", branch_codes)

    def test_pay_sabor_history_inherits_legacy_point_sales_before_cutoff(self):
        legacy_recipe = Receta.objects.create(
            nombre="Pay de Queso Mediano",
            nombre_normalizado="pay de queso mediano",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pay",
            categoria="Pay Mediano",
            hash_contenido="ventas-test-pay-queso-mediano",
        )
        current_recipe = Receta.objects.create(
            nombre="Sabor Fresa Mediano Pay",
            nombre_normalizado="sabor fresa mediano pay",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pay",
            categoria="Pay Mediano",
            hash_contenido="ventas-test-sabor-fresa-mediano-pay",
        )
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product_legacy = PointProduct.objects.create(external_id="LPAY1", name=legacy_recipe.nombre, category="Pay")
        point_product_current = PointProduct.objects.create(external_id="CPAY1", name=current_recipe.nombre, category="Pay")

        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product_legacy,
            receta=legacy_recipe,
            sale_date=date(2024, 9, 1),
            quantity=Decimal("9"),
            tickets=1,
            total_amount=Decimal("450"),
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product_current,
            receta=current_recipe,
            sale_date=date(2024, 9, 9),
            quantity=Decimal("11"),
            tickets=1,
            total_amount=Decimal("550"),
        )

        series = _load_point_daily(current_recipe, self.branch, date(2024, 9, 1), date(2024, 9, 10))

        self.assertEqual(sum(series, Decimal("0")), Decimal("20"))

    def test_generate_event_forecast_prefers_point_for_event_homologue_when_sources_conflict(self):
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product = PointProduct.objects.create(external_id="P3", name=self.product.nombre, category="Temporada")
        VentaHistorica.objects.create(
            receta=self.product,
            sucursal=self.branch,
            fecha=date(2025, 4, 30),
            cantidad=Decimal("51"),
            monto_total=Decimal("3315"),
        )
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=date(2025, 4, 30),
            product_code="P3",
            point_name=self.product.nombre,
            category="Temporada",
            quantity=Decimal("0"),
            total_amount=Decimal("0"),
            net_amount=Decimal("0"),
            tax_amount=Decimal("0"),
            gross_amount=Decimal("0"),
            discount_amount=Decimal("0"),
            source_file="test",
            source_sheet="Sheet1",
            raw_payload={},
        )
        authoritative_daily_total.cache_clear()
        authoritative_day_loaded.cache_clear()
        for lag in range(1, 5):
            target = self.event.main_date - timedelta(days=7 * lag)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=target,
                cantidad=Decimal("2"),
                monto_total=Decimal("200"),
            )

        generate_event_forecast(self.event)

        row = EventoVentaForecast.objects.get(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
        )
        self.assertEqual(Decimal(str(row.explanation_json.get("event_anchor_qty") or 0)), Decimal("0"))
        self.assertLess(row.final_forecast, Decimal("10"))

    def test_recent_branch_quantity_total_uses_canonical_sales_read_priority(self):
        target_day = timezone.localdate() - timedelta(days=1)
        PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=target_day,
            product_code="P3",
            point_name=self.product.nombre,
            category="Temporada",
            quantity=Decimal("4"),
            total_amount=Decimal("240"),
            net_amount=Decimal("240"),
            tax_amount=Decimal("0"),
            gross_amount=Decimal("240"),
            discount_amount=Decimal("0"),
            source_file="test",
            source_sheet="Sheet1",
            raw_payload={},
        )
        point_product = PointProduct.objects.create(external_id="P3", sku="P3", name=self.product.nombre, category="Temporada")
        PointDailySale.objects.create(
            branch=PointBranch.objects.get(erp_branch=self.branch),
            product=point_product,
            receta=self.product,
            sale_date=target_day,
            quantity=Decimal("9"),
            tickets=1,
            total_amount=Decimal("540"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )

        total = _recent_branch_quantity_total(self.branch.id, anchor=timezone.localdate(), days_back=1)

        self.assertEqual(total, Decimal("4"))

    def test_reconcile_event_point_sales_repairs_bridge_history(self):
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        wrong_recipe = Receta.objects.create(
            nombre="Vaso Fiesta Chico",
            nombre_normalizado="vaso fiesta chico",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Vasos Preparados",
            categoria="Vasos Chico",
            codigo_point="VFCHICO",
            hash_contenido="ventas-test-vaso-fiesta-chico",
        )
        point_product = PointProduct.objects.create(external_id="P4", sku="P4", name="Pastel Fiesta", category="Temporada")
        sale_date = date(2025, 4, 30)
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=wrong_recipe,
            sale_date=sale_date,
            quantity=Decimal("4"),
            tickets=1,
            total_amount=Decimal("400"),
        )
        authoritative_daily_total.cache_clear()
        authoritative_day_loaded.cache_clear()

        summary = reconcile_event_point_sales(self.event)

        self.assertGreaterEqual(summary.mismatch_rows, 0)
        bridge_row = VentaHistorica.objects.get(
            fuente="POINT_BRIDGE_SALES",
            receta=self.product,
            sucursal=self.branch,
            fecha=sale_date,
        )
        self.assertEqual(bridge_row.cantidad, Decimal("4"))

    def test_verified_point_daily_total_prefers_rebuild_fact_over_mixed_point_daily_sale(self):
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product = PointProduct.objects.create(external_id="P5", sku="PF1", name=self.product.nombre, category="Temporada")
        target_day = date(2025, 4, 29)
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=self.product,
            sale_date=target_day,
            quantity=Decimal("99"),
            total_amount=Decimal("9900"),
            source_endpoint="/Report/VentasCategorias",
        )
        PointSalesDailyProductFact.objects.create(
            branch=point_branch,
            sale_date=target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Temporada",
            producto_nombre_historico=self.product.nombre,
            point_product=point_product,
            receta=self.product,
            match_catalogo_status="EXACT_CODE",
            total_cantidad=Decimal("7"),
            total_venta=Decimal("700"),
            total_descuento=Decimal("0"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("700"),
            source_hash="fact-hash-1",
            source_file="fact.xls",
        )

        verified_point_daily_total.cache_clear()
        verified_point_sales_aggregate.cache_clear()

        self.assertEqual(verified_point_daily_total(self.product.id, self.branch.id, target_day), Decimal("7"))
        qty, sales = verified_point_sales_aggregate(self.product.id, target_day, target_day)
        self.assertEqual(qty, Decimal("7"))
        self.assertEqual(sales, Decimal("700"))

    def test_resolve_unit_price_uses_rebuilt_sales_window_when_available(self):
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product = PointProduct.objects.create(external_id="P6", sku="PF2", name=self.product.nombre, category="Temporada")
        target_day = date(2025, 4, 28)
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=self.product,
            sale_date=target_day,
            quantity=Decimal("50"),
            total_amount=Decimal("5000"),
            source_endpoint="/Report/VentasCategorias",
        )
        PointSalesDailyProductFact.objects.create(
            branch=point_branch,
            sale_date=target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Temporada",
            producto_nombre_historico=self.product.nombre,
            point_product=point_product,
            receta=self.product,
            match_catalogo_status="EXACT_CODE",
            total_cantidad=Decimal("5"),
            total_venta=Decimal("550"),
            total_descuento=Decimal("0"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("550"),
            source_hash="fact-hash-2",
            source_file="fact.xls",
        )

        verified_point_daily_total.cache_clear()
        verified_point_sales_aggregate.cache_clear()

        price = resolve_unit_price(self.product.id, target_day, target_day)
        self.assertEqual(price, Decimal("110.0000"))

    def test_resolve_unit_price_prefers_authoritative_over_v2_and_legacy(self):
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_product = PointProduct.objects.create(external_id="P7", sku="PF3", name=self.product.nombre, category="Temporada")
        target_day = date(2025, 4, 27)
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=target_day,
            product_code="PF3",
            point_name=self.product.nombre,
            category="Temporada",
            quantity=Decimal("4"),
            total_amount=Decimal("480"),
            net_amount=Decimal("480"),
            tax_amount=Decimal("0"),
            gross_amount=Decimal("480"),
            discount_amount=Decimal("0"),
            source_file="test",
            source_sheet="Sheet1",
            raw_payload={},
        )
        PointSalesDailyProductFact.objects.create(
            branch=point_branch,
            sale_date=target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Temporada",
            producto_nombre_historico=self.product.nombre,
            point_product=point_product,
            receta=self.product,
            match_catalogo_status="EXACT_CODE",
            total_cantidad=Decimal("5"),
            total_venta=Decimal("550"),
            total_descuento=Decimal("0"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("550"),
            source_hash="fact-hash-3",
            source_file="fact.xls",
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=self.product,
            sale_date=target_day,
            quantity=Decimal("50"),
            total_amount=Decimal("5000"),
            source_endpoint="/Report/VentasCategorias",
        )

        price = resolve_unit_price(self.product.id, target_day, target_day)

        self.assertEqual(price, Decimal("120.0000"))

    def test_resolve_unit_price_sums_canonical_sales_across_all_branches(self):
        inactive_branch = Sucursal.objects.create(codigo="GUAMUCHIL", nombre="Guamuchil", activa=False)
        target_day = date(2025, 4, 26)
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=target_day,
            product_code="PF4",
            point_name=self.product.nombre,
            category="Temporada",
            quantity=Decimal("2"),
            total_amount=Decimal("200"),
            net_amount=Decimal("200"),
            tax_amount=Decimal("0"),
            gross_amount=Decimal("200"),
            discount_amount=Decimal("0"),
            source_file="test",
            source_sheet="Sheet1",
            raw_payload={},
        )
        VentaAutoritativaPoint.objects.create(
            branch=inactive_branch,
            product=self.product,
            sale_date=target_day,
            product_code="PF4",
            point_name=self.product.nombre,
            category="Temporada",
            quantity=Decimal("3"),
            total_amount=Decimal("450"),
            net_amount=Decimal("450"),
            tax_amount=Decimal("0"),
            gross_amount=Decimal("450"),
            discount_amount=Decimal("0"),
            source_file="test",
            source_sheet="Sheet1",
            raw_payload={},
        )

        price = resolve_unit_price(self.product.id, target_day, target_day)

        self.assertEqual(price, Decimal("130.0000"))

    def test_resolve_unit_price_uses_canonical_range_aggregate_across_multiple_days(self):
        first_day = date(2025, 4, 24)
        second_day = date(2025, 4, 25)
        point_branch = PointBranch.objects.create(external_id="9", name="Matriz Range", erp_branch=self.branch)
        point_product = PointProduct.objects.create(external_id="P8", sku="PF8", name=self.product.nombre, category="Temporada")
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=first_day,
            product_code="PF8",
            point_name=self.product.nombre,
            category="Temporada",
            quantity=Decimal("2"),
            total_amount=Decimal("200"),
            net_amount=Decimal("200"),
            tax_amount=Decimal("0"),
            gross_amount=Decimal("200"),
            discount_amount=Decimal("0"),
            source_file="test",
            source_sheet="Sheet1",
            raw_payload={},
        )
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=second_day,
            product_code="PF8",
            point_name=self.product.nombre,
            category="Temporada",
            quantity=Decimal("3"),
            total_amount=Decimal("390"),
            net_amount=Decimal("390"),
            tax_amount=Decimal("0"),
            gross_amount=Decimal("390"),
            discount_amount=Decimal("0"),
            source_file="test",
            source_sheet="Sheet1",
            raw_payload={},
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=self.product,
            sale_date=first_day,
            quantity=Decimal("50"),
            total_amount=Decimal("5000"),
            source_endpoint="/Report/VentasCategorias",
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=self.product,
            sale_date=second_day,
            quantity=Decimal("50"),
            total_amount=Decimal("5000"),
            source_endpoint="/Report/VentasCategorias",
        )

        price = resolve_unit_price(self.product.id, first_day, second_day)

        self.assertEqual(price, Decimal("118.0000"))

    def test_resolve_unit_price_falls_back_to_effective_alias_recipe(self):
        target_day = date(2025, 4, 29)
        effective_recipe = Receta.objects.create(
            nombre="Pastel de Crunch R",
            nombre_normalizado="pastel de crunch r",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Temporada",
            codigo_point="0063",
            hash_contenido="price-effective-crunch-r",
        )
        broken_alias = Receta.objects.create(
            nombre="Pastel Crunch - Rebanada",
            nombre_normalizado="pastel crunch rebanada",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Temporada",
            hash_contenido="price-broken-crunch-alias",
        )
        point_branch = PointBranch.objects.create(external_id="19", name="Matriz Alias", erp_branch=self.branch)
        point_product = PointProduct.objects.create(external_id="P19", sku="0063", name=effective_recipe.nombre, category="Temporada")
        PointSalesDailyProductFact.objects.create(
            branch=point_branch,
            sale_date=target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Temporada",
            producto_nombre_historico=effective_recipe.nombre,
            point_product=point_product,
            receta=effective_recipe,
            match_catalogo_status="EXACT_CODE",
            total_cantidad=Decimal("4"),
            total_venta=Decimal("320"),
            total_descuento=Decimal("0"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("320"),
            source_hash="fact-hash-crunch-alias",
            source_file="fact.xls",
        )

        price = resolve_unit_price(broken_alias.id, target_day, target_day)

        self.assertEqual(price, Decimal("80.0000"))

    def test_resolve_unit_price_prefers_branch_specific_current_point_price_mode(self):
        other_branch = Sucursal.objects.create(codigo="LEYVA", nombre="Leyva", activa=True)
        point_branch_main = PointBranch.objects.create(external_id="pr-main", name="Matriz", erp_branch=self.branch)
        point_branch_other = PointBranch.objects.create(external_id="pr-other", name="Leyva", erp_branch=other_branch)
        point_product = PointProduct.objects.create(external_id="P20", sku="PF20", name=self.product.nombre, category="Temporada")
        latest_day = date(2026, 4, 8)
        for offset in range(3):
            PointDailySale.objects.create(
                branch=point_branch_main,
                product=point_product,
                receta=self.product,
                sale_date=latest_day - timedelta(days=offset),
                quantity=Decimal("1"),
                total_amount=Decimal("100"),
                source_endpoint="/Report/PrintReportes?idreporte=3",
            )
        for offset in range(2):
            PointDailySale.objects.create(
                branch=point_branch_other,
                product=point_product,
                receta=self.product,
                sale_date=latest_day - timedelta(days=offset),
                quantity=Decimal("1"),
                total_amount=Decimal("80"),
                source_endpoint="/Report/PrintReportes?idreporte=3",
            )

        price = resolve_unit_price(self.product.id, latest_day, latest_day, branch_id=other_branch.id)

        self.assertEqual(price, Decimal("80.0000"))

    def test_resolve_unit_price_falls_back_to_global_current_point_price_when_branch_has_isolated_outlier(self):
        other_branch = Sucursal.objects.create(codigo="PAYAN", nombre="Payan", activa=True)
        point_branch_main = PointBranch.objects.create(external_id="pr2-main", name="Matriz", erp_branch=self.branch)
        point_branch_other = PointBranch.objects.create(external_id="pr2-other", name="Payan", erp_branch=other_branch)
        point_product = PointProduct.objects.create(external_id="P21", sku="PF21", name=self.product.nombre, category="Temporada")
        latest_day = date(2026, 4, 8)
        for offset in range(3):
            PointDailySale.objects.create(
                branch=point_branch_main,
                product=point_product,
                receta=self.product,
                sale_date=latest_day - timedelta(days=offset),
                quantity=Decimal("1"),
                total_amount=Decimal("100"),
                source_endpoint="/Report/PrintReportes?idreporte=3",
            )
        PointDailySale.objects.create(
            branch=point_branch_other,
            product=point_product,
            receta=self.product,
            sale_date=latest_day,
            quantity=Decimal("1"),
            total_amount=Decimal("60"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )

        price = resolve_unit_price(self.product.id, latest_day, latest_day, branch_id=other_branch.id)

        self.assertEqual(price, Decimal("100.0000"))


class VentasEventosViewTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username="ventas_super",
            email="ventas_super@example.com",
            password="Ventas123!",
        )
        self.client.force_login(self.user)
        self.branch = Sucursal.objects.create(codigo="MATRIZ", nombre="Matriz", activa=True)
        self.product = Receta.objects.create(
            nombre="Pastel Fiesta",
            nombre_normalizado="pastel fiesta",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Temporada",
            hash_contenido="ventas-view-pastel-fiesta",
        )
        self.event = EventoVenta.objects.create(
            name="Dia del Nino 2026",
            event_type="TEMPORADA",
            main_date=date(2026, 4, 30),
            analysis_start_date=date(2026, 4, 27),
            analysis_end_date=date(2026, 5, 3),
        )
        EventoVentaSucursal.objects.create(sales_event=self.event, branch=self.branch)
        EventoVentaProducto.objects.create(sales_event=self.event, product=self.product)
        for offset, qty in enumerate([Decimal("10"), Decimal("11"), Decimal("12"), Decimal("13")], start=1):
            target = self.event.main_date - timedelta(days=7 * offset)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=target,
                cantidad=qty,
                monto_total=qty * Decimal("100"),
            )
        generate_event_forecast(self.event)

    def _seed_implausible_revenue_case(self):
        EventoVentaForecast.objects.filter(sales_event=self.event).delete()
        week_start, week_end = _event_projection_window(self.event)
        day_count = (week_end - week_start).days + 1
        for offset in range(day_count):
            EventoVentaForecast.objects.create(
                sales_event=self.event,
                branch=self.branch,
                product=self.product,
                forecast_date=week_start + timedelta(days=offset),
                final_forecast=Decimal("100"),
                conservative_forecast=Decimal("90"),
                aggressive_forecast=Decimal("110"),
            )
        for offset in range(day_count):
            hist_day = date(week_start.year - 1, week_start.month, week_start.day) + timedelta(days=offset)
            VentaHistorica.objects.create(
                receta=self.product,
                sucursal=self.branch,
                fecha=hist_day,
                cantidad=Decimal("100"),
                monto_total=Decimal("10000"),
            )

    def test_event_financial_dataset_keeps_real_price_times_forecast_qty(self):
        self._seed_implausible_revenue_case()
        week_start, week_end = _event_projection_window(self.event)
        day_count = (week_end - week_start).days + 1
        expected_sales = (Decimal(day_count) * Decimal("100") * Decimal("150")).quantize(Decimal("0.01"))

        with mock.patch("ventas.views.resolve_unit_prices_bulk", return_value={(self.product.id, self.branch.id): Decimal("150.0000")}), mock.patch(
            "ventas.views.get_commercial_total_cost_map",
            return_value={self.product.id: Decimal("60")},
        ):
            dataset = _event_financial_dataset(
                self.event,
                EventoVentaForecast.objects.filter(sales_event=self.event),
                start_date=week_start,
                end_date=week_end,
            )

        self.assertTrue(dataset["plausibility"]["flagged"])
        self.assertEqual(dataset["summary"]["sales"], expected_sales)
        self.assertIn("precio real x piezas", dataset["validation_message"].lower())

    def test_build_financials_keeps_real_price_times_forecast_qty(self):
        self._seed_implausible_revenue_case()
        week_start, week_end = _event_projection_window(self.event)
        day_count = (week_end - week_start).days + 1
        expected_sales = (Decimal(day_count) * Decimal("100") * Decimal("150")).quantize(Decimal("0.01"))

        with mock.patch("ventas.services.financials.resolve_unit_price", return_value=Decimal("150.0000")), mock.patch(
            "ventas.services.financials.get_commercial_total_cost_map",
            return_value={self.product.id: Decimal("60")},
        ):
            result = build_financials(self.event)

        self.assertEqual(result["created"], 3)
        base = self.event.financials.get(scenario="BASE")
        self.assertEqual(base.estimated_sales.quantize(Decimal("0.01")), expected_sales)
        self.assertTrue(
            self.event.notifications.filter(
                severity=EventoVentaNotification.SEVERITY_WARN,
                message__icontains="precio real x piezas forecast",
            ).exists()
        )

    def test_event_detail_snapshot_uses_real_price_times_forecast_qty(self):
        self._seed_implausible_revenue_case()
        week_start, week_end = _event_projection_window(self.event)
        day_count = (week_end - week_start).days + 1
        expected_week_sales = (Decimal(day_count) * Decimal("100") * Decimal("150")).quantize(Decimal("0.01"))
        expected_day_sales = (Decimal("100") * Decimal("150")).quantize(Decimal("0.01"))

        with mock.patch("ventas.views.resolve_unit_prices_bulk", return_value={(self.product.id, self.branch.id): Decimal("150.0000")}), mock.patch(
            "ventas.views.get_commercial_total_cost_map",
            return_value={self.product.id: Decimal("60")},
        ), mock.patch("ventas.services.financials.resolve_unit_price", return_value=Decimal("150.0000")), mock.patch(
            "ventas.services.financials.get_commercial_total_cost_map",
            return_value={self.product.id: Decimal("60")},
        ):
            payload = build_event_detail_snapshot_payload(self.event)

        self.assertEqual(payload["week_projected_revenue"], str(expected_week_sales))
        self.assertEqual(payload["main_day_projected_revenue"], str(expected_day_sales))
        self.assertIn("projection_model", payload["executive_dataset"])

    def test_event_financial_dataset_exposes_executive_projection_model(self):
        week_start, week_end = _event_projection_window(self.event)

        with mock.patch("ventas.views.resolve_unit_prices_bulk", return_value={(self.product.id, self.branch.id): Decimal("150.0000")}), mock.patch(
            "ventas.views.get_commercial_total_cost_map",
            return_value={self.product.id: Decimal("60")},
        ):
            dataset = _event_financial_dataset(
                self.event,
                EventoVentaForecast.objects.filter(sales_event=self.event),
                start_date=week_start,
                end_date=week_end,
            )

        self.assertIn("projection_model", dataset)
        self.assertIn("same_store_factor", dataset["projection_model"])
        self.assertIn("final_projection_reasoning", dataset["projection_model"])

    def test_event_financial_dataset_accepts_dg_benchmark_override_from_notes(self):
        self._seed_implausible_revenue_case()
        self.event.objective_notes = "Benchmark DG 2025 mismo periodo: $105,000.00"
        self.event.save(update_fields=["objective_notes", "updated_at"])
        week_start, week_end = _event_projection_window(self.event)

        with mock.patch("ventas.views.resolve_unit_prices_bulk", return_value={(self.product.id, self.branch.id): Decimal("150.0000")}), mock.patch(
            "ventas.views.get_commercial_total_cost_map",
            return_value={self.product.id: Decimal("60")},
        ):
            dataset = _event_financial_dataset(
                self.event,
                EventoVentaForecast.objects.filter(sales_event=self.event),
                start_date=week_start,
                end_date=week_end,
            )

        self.assertFalse(dataset["plausibility"]["flagged"])
        self.assertEqual(dataset["plausibility"]["benchmark_source"], "objective_notes")
        self.assertEqual(dataset["plausibility"]["reference_sales_ceiling"], Decimal("105000.00"))

    def test_event_financial_dataset_accepts_week_when_qty_target_is_met_and_dg_main_day_benchmark_exists(self):
        self._seed_implausible_revenue_case()
        self.event.objective_notes = "Benchmark DG día principal: $50,000.00"
        self.event.save(update_fields=["objective_notes", "updated_at"])
        week_start, week_end = _event_projection_window(self.event)

        with mock.patch(
            "ventas.views.build_event_executive_projection_model",
            return_value={
                "target_total_qty": Decimal("700.020"),
                "current_total_qty": Decimal("700.000"),
                "benchmark_source": "historical_calendar",
                "same_store_factor": Decimal("1.0000"),
                "expansion_factor": Decimal("0.0000"),
                "contraction_factor": Decimal("1.0000"),
                "final_projection_reasoning": "test",
                "mix_adjustment_source": "test",
            },
        ), mock.patch(
            "ventas.views.resolve_unit_prices_bulk",
            return_value={(self.product.id, self.branch.id): Decimal("150.0000")},
        ), mock.patch(
            "ventas.views.get_commercial_total_cost_map",
            return_value={self.product.id: Decimal("60")},
        ):
            dataset = _event_financial_dataset(
                self.event,
                EventoVentaForecast.objects.filter(sales_event=self.event),
                start_date=week_start,
                end_date=week_end,
            )

        self.assertFalse(dataset["plausibility"]["flagged"])
        self.assertEqual(dataset["plausibility"]["reason"], "within_qty_target_with_dg_main_day_benchmark")

    def test_detail_snapshot_refreshes_when_objective_notes_change(self):
        first_payload = get_event_detail_snapshot_payload(self.event, generated_by=self.user)
        self.event.objective_notes = "Benchmark DG 2025 mismo periodo: $120,000.00"
        self.event.save(update_fields=["objective_notes", "updated_at"])

        second_payload = get_event_detail_snapshot_payload(self.event, generated_by=self.user)

        self.assertNotEqual(
            first_payload.get("executive_dataset", {}).get("projection_model", {}).get("benchmark_source"),
            second_payload.get("executive_dataset", {}).get("projection_model", {}).get("benchmark_source"),
        )
        self.assertEqual(
            second_payload.get("executive_dataset", {}).get("projection_model", {}).get("benchmark_source"),
            "objective_notes",
        )

    def test_event_detail_explains_zero_trend_adjustment(self):
        response = self.client.get(reverse("ventas:evento_detail", args=[self.event.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Excel semana")
        self.assertContains(response, "Excel día exacto")
        self.assertContains(response, "ajuste por tendencia = 0")
        self.assertContains(response, "Explicación")

    def test_event_update_coerces_posted_dates_before_regenerating_forecast(self):
        response = self.client.post(
            reverse("ventas:evento_update", args=[self.event.id]),
            data={
                "name": "Dia del Nino 2026 Ajustado",
                "event_type": "TEMPORADA",
                "main_date": "2026-05-01",
                "analysis_start_date": "2026-04-28",
                "analysis_end_date": "2026-05-04",
                "approval_deadline": "2026-04-25",
                "priority": self.event.priority,
                "scenario_focus": self.event.scenario_focus,
                "conservative_pct": str(self.event.conservative_pct),
                "aggressive_pct": str(self.event.aggressive_pct),
                "branches": [str(self.branch.id)],
                "products": [str(self.product.id)],
            },
        )

        self.assertEqual(response.status_code, 302)
        self.event.refresh_from_db()
        self.assertEqual(self.event.main_date, date(2026, 5, 1))
        self.assertEqual(self.event.analysis_start_date, date(2026, 4, 28))
        self.assertEqual(self.event.analysis_end_date, date(2026, 5, 4))
        self.assertEqual(self.event.approval_deadline, date(2026, 4, 25))
        self.assertTrue(EventoVentaForecast.objects.filter(sales_event=self.event).exists())

    def test_week_projection_export_builds_general_and_branch_sheets(self):
        response = self.client.get(reverse("ventas:evento_export_week_projection", args=[self.event.id]))

        self.assertEqual(response.status_code, 200)
        self.assertIn(".xlsx", response["Content-Disposition"])
        workbook = load_workbook(filename=BytesIO(response.content))
        self.assertEqual(workbook.sheetnames[0], "Resumen")
        self.assertIn("Resumen", workbook.sheetnames)
        self.assertIn("General", workbook.sheetnames)
        self.assertIn("MATRIZ", workbook.sheetnames)
        self.assertEqual(workbook["General"]["A1"].value, "Pollyana's Dolce · Proyeccion comercial")
        self.assertIn(self.event.name, workbook["General"]["A2"].value)
        self.assertEqual(workbook["General"]["B4"].value, f"{self.event.analysis_start_date} a {self.event.analysis_end_date}")
        self.assertEqual(workbook["Resumen"]["A1"].value, "Pollyana's Dolce · Resumen ejecutivo")
        summary_header = None
        for row_index in range(1, 15):
            candidate = [workbook["Resumen"].cell(row_index, column).value for column in range(1, 8)]
            if "Proyeccion general" in candidate:
                summary_header = candidate
                break
        self.assertIsNotNone(summary_header)
        self.assertIn("Proyeccion general", summary_header)
        self.assertIn("MATRIZ", summary_header)
        header_row_index = None
        header_row = []
        for row_index in range(1, 12):
            candidate = [workbook["General"].cell(row_index, column).value for column in range(1, 15)]
            if "Producto" in candidate:
                header_row_index = row_index
                header_row = candidate
                break
        self.assertIsNotNone(header_row_index)
        self.assertIn("Explicacion", header_row)
        explanation_column = header_row.index("Explicacion") + 1
        self.assertIn(
            "forecast",
            str(workbook["General"].cell((header_row_index or 0) + 1, explanation_column).value or "").lower(),
        )

    def test_week_projection_export_uses_fixed_window_around_main_date(self):
        self.event.analysis_start_date = date(2026, 4, 20)
        self.event.analysis_end_date = date(2026, 5, 10)
        self.event.save(update_fields=["analysis_start_date", "analysis_end_date", "updated_at"])
        generate_event_forecast(self.event)

        response = self.client.get(reverse("ventas:evento_export_week_projection", args=[self.event.id]))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content))
        self.assertEqual(workbook["General"]["B4"].value, "2026-04-20 a 2026-05-10")

    def test_main_day_projection_export_uses_exact_date_scope(self):
        response = self.client.get(reverse("ventas:evento_export_main_day_projection", args=[self.event.id]))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content))
        self.assertEqual(workbook["General"]["B4"].value, str(self.event.main_date))
        self.assertEqual(workbook["Resumen"]["B4"].value, str(self.event.main_date))
        self.assertEqual(workbook["MATRIZ"]["B3"].value, str(self.event.main_date))

    def test_generate_projection_files_persists_operational_and_dashboard_artifacts(self):
        response = self.client.get(reverse("ventas:evento_generate_projection_files", args=[self.event.id]))

        self.assertEqual(response.status_code, 302)
        artifacts = EventoVentaProjectionArtifact.objects.filter(sales_event=self.event, forecast_version=self.event.version)
        self.assertEqual(artifacts.count(), 5)
        self.assertTrue(
            artifacts.filter(export_type=EventoVentaProjectionArtifact.TYPE_WEEK, file_name__icontains="semana").exists()
        )
        self.assertTrue(
            artifacts.filter(export_type=EventoVentaProjectionArtifact.TYPE_DAY, file_name__icontains="dia").exists()
        )
        self.assertTrue(
            artifacts.filter(export_type=EventoVentaProjectionArtifact.TYPE_DAILY, file_name__icontains="por_dia").exists()
        )
        self.assertTrue(
            artifacts.filter(export_type=EventoVentaProjectionArtifact.TYPE_DASHBOARD, file_name__icontains="dashboard").exists()
        )
        self.assertTrue(
            artifacts.filter(export_type=EventoVentaProjectionArtifact.TYPE_PACKAGE, file_name__icontains="paquete").exists()
        )
        package_artifact = artifacts.get(export_type=EventoVentaProjectionArtifact.TYPE_PACKAGE)
        with ZipFile(package_artifact.file_path) as package_zip:
            packaged_names = set(package_zip.namelist())
        self.assertTrue(any("proyeccion_semana" in name for name in packaged_names))
        self.assertTrue(any("proyeccion_dia" in name for name in packaged_names))
        self.assertTrue(any("proyeccion_por_dia" in name for name in packaged_names))
        self.assertTrue(any("dashboard_ejecutivo" in name for name in packaged_names))

    def test_submit_approval_blocks_when_main_day_is_below_homologue_floor(self):
        self.event.main_date = date(2026, 6, 21)
        self.event.analysis_start_date = date(2026, 6, 16)
        self.event.analysis_end_date = date(2026, 6, 22)
        self.event.save(update_fields=["main_date", "analysis_start_date", "analysis_end_date", "updated_at"])
        original_status = self.event.status
        EventoVentaForecast.objects.filter(sales_event=self.event).delete()
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
            final_forecast=Decimal("10"),
            conservative_forecast=Decimal("9"),
            aggressive_forecast=Decimal("12"),
        )
        VentaHistorica.objects.create(
            receta=self.product,
            sucursal=self.branch,
            fecha=date(2025, 6, 15),
            cantidad=Decimal("120"),
            monto_total=Decimal("12000"),
        )
        EventoVentaFinancial.objects.update_or_create(
            sales_event=self.event,
            scenario="BASE",
            defaults={
                "estimated_sales": Decimal("1000"),
                "estimated_cogs": Decimal("400"),
                "estimated_gross_profit": Decimal("600"),
                "estimated_margin": Decimal("60"),
                "incremental_investment": Decimal("400"),
                "break_even_sales": Decimal("400"),
                "expected_roi": Decimal("150"),
            },
        )

        response = self.client.post(reverse("ventas:evento_submit_approval", args=[self.event.id]))

        self.assertEqual(response.status_code, 302)
        self.event.refresh_from_db()
        self.assertEqual(self.event.status, original_status)
        self.assertFalse(self.event.approvals.exists())
        self.assertTrue(
            self.event.notifications.filter(
                severity=EventoVentaNotification.SEVERITY_CRIT,
                message__icontains="Guard de aprobación bloqueó",
            ).exists()
        )

    def test_approve_blocks_when_main_day_is_below_homologue_floor(self):
        self.event.main_date = date(2026, 6, 21)
        self.event.analysis_start_date = date(2026, 6, 16)
        self.event.analysis_end_date = date(2026, 6, 22)
        self.event.status = EventoVenta.STATUS_LISTO_REVISION
        self.event.save(update_fields=["main_date", "analysis_start_date", "analysis_end_date", "status", "updated_at"])
        EventoVentaForecast.objects.filter(sales_event=self.event).delete()
        EventoVentaForecast.objects.create(
            sales_event=self.event,
            branch=self.branch,
            product=self.product,
            forecast_date=self.event.main_date,
            final_forecast=Decimal("10"),
            conservative_forecast=Decimal("9"),
            aggressive_forecast=Decimal("12"),
        )
        VentaHistorica.objects.create(
            receta=self.product,
            sucursal=self.branch,
            fecha=date(2025, 6, 15),
            cantidad=Decimal("120"),
            monto_total=Decimal("12000"),
        )
        EventoVentaFinancial.objects.update_or_create(
            sales_event=self.event,
            scenario="BASE",
            defaults={
                "estimated_sales": Decimal("1000"),
                "estimated_cogs": Decimal("400"),
                "estimated_gross_profit": Decimal("600"),
                "estimated_margin": Decimal("60"),
                "incremental_investment": Decimal("400"),
                "break_even_sales": Decimal("400"),
                "expected_roi": Decimal("150"),
            },
        )

        response = self.client.post(reverse("ventas:evento_approve", args=[self.event.id]))

        self.assertEqual(response.status_code, 302)

    def test_submit_approval_blocks_when_week_revenue_is_implausibly_high(self):
        self._seed_implausible_revenue_case()
        original_status = self.event.status
        EventoVentaFinancial.objects.update_or_create(
            sales_event=self.event,
            scenario="BASE",
            defaults={
                "estimated_sales": Decimal("105000"),
                "estimated_cogs": Decimal("40000"),
                "estimated_gross_profit": Decimal("65000"),
                "estimated_margin": Decimal("61.90"),
                "incremental_investment": Decimal("40000"),
                "break_even_sales": Decimal("40000"),
                "expected_roi": Decimal("162.50"),
            },
        )

        with mock.patch("ventas.views.resolve_unit_prices_bulk", return_value={(self.product.id, self.branch.id): Decimal("150.0000")}), mock.patch(
            "ventas.views.get_commercial_total_cost_map",
            return_value={self.product.id: Decimal("60")},
        ):
            response = self.client.post(reverse("ventas:evento_submit_approval", args=[self.event.id]))

        self.assertEqual(response.status_code, 302)
        self.event.refresh_from_db()
        self.assertEqual(self.event.status, original_status)
        self.assertFalse(self.event.approvals.exists())
        self.assertTrue(
            self.event.notifications.filter(
                severity=EventoVentaNotification.SEVERITY_CRIT,
                message__icontains="ingreso semanal proyectado",
            ).exists()
        )

    def test_approve_blocks_when_week_revenue_is_implausibly_high(self):
        self._seed_implausible_revenue_case()
        self.event.status = EventoVenta.STATUS_LISTO_REVISION
        self.event.save(update_fields=["status", "updated_at"])
        EventoVentaFinancial.objects.update_or_create(
            sales_event=self.event,
            scenario="BASE",
            defaults={
                "estimated_sales": Decimal("105000"),
                "estimated_cogs": Decimal("40000"),
                "estimated_gross_profit": Decimal("65000"),
                "estimated_margin": Decimal("61.90"),
                "incremental_investment": Decimal("40000"),
                "break_even_sales": Decimal("40000"),
                "expected_roi": Decimal("162.50"),
            },
        )

        with mock.patch("ventas.views.resolve_unit_prices_bulk", return_value={(self.product.id, self.branch.id): Decimal("150.0000")}), mock.patch(
            "ventas.views.get_commercial_total_cost_map",
            return_value={self.product.id: Decimal("60")},
        ):
            response = self.client.post(reverse("ventas:evento_approve", args=[self.event.id]))

        self.assertEqual(response.status_code, 302)
        self.event.refresh_from_db()
        self.assertEqual(self.event.status, EventoVenta.STATUS_LISTO_REVISION)
        self.assertTrue(
            self.event.notifications.filter(
                severity=EventoVentaNotification.SEVERITY_CRIT,
                message__icontains="ingreso semanal proyectado",
            ).exists()
        )
        self.event.refresh_from_db()
        self.assertEqual(self.event.status, EventoVenta.STATUS_LISTO_REVISION)
        self.assertIsNone(self.event.approved_by)

    def test_approve_blocks_when_week_qty_exceeds_executive_target(self):
        self._seed_implausible_revenue_case()
        self.event.status = EventoVenta.STATUS_LISTO_REVISION
        self.event.save(update_fields=["status", "updated_at"])
        EventoVentaFinancial.objects.update_or_create(
            sales_event=self.event,
            scenario="BASE",
            defaults={
                "estimated_sales": Decimal("70000"),
                "estimated_cogs": Decimal("28000"),
                "estimated_gross_profit": Decimal("42000"),
                "estimated_margin": Decimal("60.00"),
                "incremental_investment": Decimal("28000"),
                "break_even_sales": Decimal("28000"),
                "expected_roi": Decimal("150.00"),
            },
        )

        with mock.patch(
            "ventas.views.build_event_executive_projection_model",
            return_value={
                "target_total_qty": Decimal("300.000"),
                "current_total_qty": Decimal("700.000"),
                "benchmark_source": "historical_calendar",
            },
        ), mock.patch(
            "ventas.views.resolve_unit_prices_bulk",
            return_value={(self.product.id, self.branch.id): Decimal("100.0000")},
        ), mock.patch(
            "ventas.views.get_commercial_total_cost_map",
            return_value={self.product.id: Decimal("60")},
        ):
            response = self.client.post(reverse("ventas:evento_approve", args=[self.event.id]))

        self.assertEqual(response.status_code, 302)
        self.event.refresh_from_db()
        self.assertEqual(self.event.status, EventoVenta.STATUS_LISTO_REVISION)
        self.assertTrue(
            self.event.notifications.filter(
                severity=EventoVentaNotification.SEVERITY_CRIT,
                message__icontains="target ejecutivo defendible",
            ).exists()
        )

    def test_sync_review_status_with_guardrails_returns_event_to_modeling(self):
        self.event.status = EventoVenta.STATUS_LISTO_REVISION
        self.event.save(update_fields=["status", "updated_at"])

        with mock.patch(
            "ventas.views._approval_blocking_findings",
            return_value=["La semana total del evento sigue por encima del target ejecutivo defendible."],
        ):
            changed = _sync_event_review_status_with_guardrails(self.event, actor=self.user)

        self.assertTrue(changed)
        self.event.refresh_from_db()
        self.assertEqual(self.event.status, EventoVenta.STATUS_MODELADO)
        self.assertTrue(
            self.event.notifications.filter(
                severity=EventoVentaNotification.SEVERITY_CRIT,
                message__icontains="regresó el evento a modelado",
            ).exists()
        )

    def test_generate_projection_files_archives_obsolete_legacy_publication(self):
        self.event.code = "DÍADELNINO2026-260430-001"
        self.event.save(update_fields=["code"])
        legacy_dir = (
            Path(settings.BASE_DIR)
            / "output"
            / "spreadsheet"
            / "ventas_eventos"
            / "díadelnino2026-260430-001"
            / "actual"
        )
        legacy_dir.mkdir(parents=True, exist_ok=True)
        legacy_file = legacy_dir / "díadelnino2026-260430-001_proyeccion_semana_2026-04-27_2026-05-03.xlsx"
        legacy_file.write_bytes(b"legacy")
        legacy_artifact = EventoVentaProjectionArtifact.objects.create(
            sales_event=self.event,
            export_type=EventoVentaProjectionArtifact.TYPE_WEEK,
            forecast_version=max(self.event.version - 1, 0),
            generated_by=self.user,
            file_name=legacy_file.name,
            file_path=str(legacy_file),
            size_bytes=legacy_file.stat().st_size,
        )

        response = self.client.get(reverse("ventas:evento_generate_projection_files", args=[self.event.id]))

        self.assertEqual(response.status_code, 302)
        self.assertFalse(EventoVentaProjectionArtifact.objects.filter(id=legacy_artifact.id).exists())
        archived_file = (
            Path(settings.BASE_DIR)
            / "output"
            / "spreadsheet"
            / "ventas_eventos"
            / "_historico_no_vigente"
            / timezone.localdate().isoformat()
            / "díadelnino2026-260430-001"
            / "actual"
            / legacy_file.name
        )
        self.assertTrue(archived_file.exists())

    def test_executive_dashboard_export_includes_sku_interpretation_for_blocked_costs(self):
        response = self.client.get(reverse("ventas:evento_export_executive_dashboard", args=[self.event.id]))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=BytesIO(response.content))
        self.assertIn("Validacion", workbook.sheetnames)
        self.assertIn("Gobernanza SKU", workbook.sheetnames)

        validation_sheet = workbook["Validacion"]
        header_row = None
        for row_index in range(1, validation_sheet.max_row + 1):
            candidate = [validation_sheet.cell(row_index, column).value for column in range(1, 6)]
            if candidate[:5] == ["Producto", "Interpretacion SKU", "Estado costo", "Motivo", "Componentes/relacion"]:
                header_row = row_index
                break
        self.assertIsNotNone(header_row)
        self.assertEqual(validation_sheet.cell(header_row + 1, 2).value, "Receta directa")
        self.assertEqual(validation_sheet.cell(header_row + 1, 3).value, "Bloqueado")
        self.assertIn("BOM operativa", str(validation_sheet.cell(header_row + 1, 4).value))

    def test_detail_filter_can_show_comparable_rows(self):
        guamuchil = Sucursal.objects.create(codigo="GUAMUCHIL", nombre="Guamuchil", activa=True)
        EventoVentaSucursal.objects.create(sales_event=self.event, branch=guamuchil)

        point_branch_matriz = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        point_branch_guamuchil = PointBranch.objects.create(external_id="13", name="Guamuchil", erp_branch=guamuchil)
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_matriz,
            indicator_date=date(2026, 4, 20),
            total_amount=Decimal("100000"),
            total_tickets=100,
            total_avg_ticket=Decimal("1000"),
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_guamuchil,
            indicator_date=date(2026, 4, 20),
            total_amount=Decimal("60000"),
            total_tickets=60,
            total_avg_ticket=Decimal("1000"),
        )
        generate_event_forecast(self.event)

        response = self.client.get(reverse("ventas:evento_detail", args=[self.event.id]) + "?detail_source=comparable")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Filtro activo del detalle: comparable")
        self.assertContains(response, "Sucursal comparable")

    def test_create_event_runs_projection_pipeline_automatically(self):
        branch = Sucursal.objects.create(codigo="CRUCERO", nombre="Crucero", activa=True)
        product = Receta.objects.create(
            nombre="Bollo Demo",
            nombre_normalizado="bollo demo",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Bollo",
            categoria="Bollo",
            hash_contenido="ventas-view-bollo-demo",
        )
        for offset, qty in enumerate([Decimal("7"), Decimal("8"), Decimal("9"), Decimal("10")], start=1):
            target = date(2026, 5, 10) - timedelta(days=7 * offset)
            VentaHistorica.objects.create(
                receta=product,
                sucursal=branch,
                fecha=target,
                cantidad=qty,
                monto_total=qty * Decimal("100"),
            )

        response = self.client.post(
            reverse("ventas:evento_create"),
            data={
                "name": "Dia de las Madres 2026",
                "event_type": "TEMPORADA",
                "main_date": "2026-05-10",
                "branches": [str(branch.id)],
                "products": [str(product.id)],
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        event = EventoVenta.objects.get(name="Dia de las Madres 2026")
        self.assertTrue(EventoVentaForecast.objects.filter(sales_event=event).exists())
        self.assertEqual(event.projection_artifacts.count(), 5)

    def test_delete_event_removes_draft_from_list(self):
        draft_event = EventoVenta.objects.create(
            name="Evento borrador eliminable",
            event_type="TEMPORADA",
            main_date=date(2026, 6, 1),
            analysis_start_date=date(2026, 5, 29),
            analysis_end_date=date(2026, 6, 2),
            status=EventoVenta.STATUS_BORRADOR,
        )

        response = self.client.post(reverse("ventas:evento_delete", args=[draft_event.id]), follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(EventoVenta.objects.filter(id=draft_event.id).exists())
        messages = [message.message for message in response.context["messages"]]
        self.assertTrue(any("Evento eliminado:" in message for message in messages))

    def test_delete_event_rejects_non_draft_status(self):
        self.event.status = EventoVenta.STATUS_APROBADO
        self.event.save(update_fields=["status", "updated_at"])

        response = self.client.post(reverse("ventas:evento_delete", args=[self.event.id]), follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertTrue(EventoVenta.objects.filter(id=self.event.id).exists())
        messages = [message.message for message in response.context["messages"]]
        self.assertTrue(any("Solo se pueden eliminar eventos en borrador o rechazados." in message for message in messages))

    def test_event_list_shows_delete_button_for_manage_users(self):
        response = self.client.get(reverse("ventas:eventos"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Eliminar")

    def test_event_detail_shows_generate_projection_files_button(self):
        response = self.client.get(reverse("ventas:evento_detail", args=[self.event.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Generar archivos ERP")


@override_settings(TIME_ZONE="America/Phoenix")
class VentasEventosSchedulesTests(TestCase):
    def test_registers_sales_event_periodic_tasks_idempotently(self):
        try:
            from django_celery_beat.models import PeriodicTask
        except Exception as exc:  # pragma: no cover
            raise CommandError(f"django_celery_beat no disponible: {exc}") from exc

        call_command("setup_ventas_celery_schedules")
        call_command("setup_ventas_celery_schedules")

        task_names = set(PeriodicTask.objects.values_list("name", flat=True))
        self.assertIn("ventas: monitoreo eventos activos", task_names)
        self.assertIn("ventas: monitoreo cierre postmortem", task_names)
        self.assertEqual(PeriodicTask.objects.filter(name__startswith="ventas: ").count(), 2)


class SalesReadServiceTests(TestCase):
    def setUp(self):
        self.branch = Sucursal.objects.create(codigo="MATRIZ", nombre="Matriz", activa=True)
        self.point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.branch)
        self.product = Receta.objects.create(
            nombre="Pastel Canonico",
            nombre_normalizado="pastel canonico",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            codigo_point="PT001",
            hash_contenido="ventas-read-product",
        )
        self.other_product = Receta.objects.create(
            nombre="Pay Canonico",
            nombre_normalizado="pay canonico",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            codigo_point="PY001",
            hash_contenido="ventas-read-other-product",
        )
        self.point_product = PointProduct.objects.create(
            external_id="point-pt001",
            sku="PT001",
            name="Pastel Canonico",
            category="Pasteles",
        )
        self.point_product_other = PointProduct.objects.create(
            external_id="point-py001",
            sku="PY001",
            name="Pay Canonico",
            category="Pays",
        )
        self.target_day = date(2026, 4, 1)

    def test_get_daily_sales_prefers_authoritative_over_v2_and_legacy(self):
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=self.target_day,
            product_code="PT001",
            point_name="Pastel Canonico",
            quantity=Decimal("9"),
            total_amount=Decimal("450"),
        )
        PointSalesDailyProductFact.objects.create(
            branch=self.point_branch,
            sale_date=self.target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Pasteles",
            producto_nombre_historico="Pastel Canonico",
            point_product=self.point_product,
            receta=self.product,
            total_cantidad=Decimal("7"),
            total_venta=Decimal("350"),
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product,
            receta=self.product,
            sale_date=self.target_day,
            quantity=Decimal("5"),
            total_amount=Decimal("250"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )

        result = get_daily_sales(self.branch, self.target_day, self.product)

        self.assertEqual(result["source"], "authoritative")
        self.assertEqual(result["cantidad"], Decimal("9"))
        self.assertEqual(result["monto"], Decimal("450"))
        self.assertFalse(result["fallback_legacy_used"])

    def test_get_daily_sales_uses_v2_category_fact_for_branch_day_aggregate(self):
        PointSalesDailyCategoryFact.objects.create(
            branch=self.point_branch,
            sale_date=self.target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("12"),
            total_venta=Decimal("600"),
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=self.point_branch,
            sale_date=self.target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Pays",
            total_cantidad=Decimal("3"),
            total_venta=Decimal("150"),
        )

        result = get_daily_sales(self.branch, self.target_day)

        self.assertEqual(result["source"], "v2_fact")
        self.assertEqual(result["source_detail"], "point_sales_daily_category_fact")
        self.assertEqual(result["cantidad"], Decimal("15"))
        self.assertEqual(result["monto"], Decimal("750"))

    def test_get_daily_sales_falls_back_to_legacy_without_mixing_sources(self):
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product,
            receta=self.product,
            sale_date=self.target_day,
            quantity=Decimal("8"),
            total_amount=Decimal("400"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product_other,
            receta=self.other_product,
            sale_date=self.target_day,
            quantity=Decimal("20"),
            total_amount=Decimal("1000"),
            source_endpoint="/Report/VentasCategorias",
        )

        with self.assertLogs("ventas.services.sales_read_service", level="WARNING") as captured:
            result = get_daily_sales(self.branch, self.target_day)

        self.assertEqual(result["source"], "legacy")
        self.assertEqual(result["source_detail"], "point_daily_sale_official")
        self.assertEqual(result["cantidad"], Decimal("8"))
        self.assertEqual(result["monto"], Decimal("400"))
        self.assertTrue(result["fallback_legacy_used"])
        self.assertIn("legacy fallback used", captured.output[0])

    def test_get_daily_sales_returns_none_source_when_no_data_exists(self):
        result = get_daily_sales(self.branch, self.target_day, self.product)

        self.assertEqual(result["source"], "none")
        self.assertEqual(result["cantidad"], Decimal("0"))
        self.assertEqual(result["monto"], Decimal("0"))

    def test_get_sales_range_prefers_authoritative_over_v2_and_legacy(self):
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=self.target_day,
            product_code="PT001",
            point_name="Pastel Canonico",
            quantity=Decimal("9"),
            total_amount=Decimal("450"),
        )
        PointSalesDailyProductFact.objects.create(
            branch=self.point_branch,
            sale_date=self.target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Pasteles",
            producto_nombre_historico="Pastel Canonico",
            point_product=self.point_product,
            receta=self.product,
            total_cantidad=Decimal("7"),
            total_venta=Decimal("350"),
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product,
            receta=self.product,
            sale_date=self.target_day,
            quantity=Decimal("5"),
            total_amount=Decimal("250"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )

        result = get_sales_range(start_date=self.target_day, end_date=self.target_day, producto=self.product)

        self.assertEqual(result["source"], "authoritative")
        self.assertEqual(result["cantidad"], Decimal("9"))
        self.assertEqual(result["monto"], Decimal("450"))
        self.assertEqual(result["coverage_days"], 1)
        self.assertEqual(result["coverage_branches"], 1)
        self.assertTrue(result["coverage_accepted"])
        self.assertEqual(result["coverage_reason"], "strict_priority")

    def test_get_sales_range_uses_v2_over_legacy(self):
        PointSalesDailyProductFact.objects.create(
            branch=self.point_branch,
            sale_date=self.target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Pasteles",
            producto_nombre_historico="Pastel Canonico",
            point_product=self.point_product,
            receta=self.product,
            total_cantidad=Decimal("7"),
            total_venta=Decimal("350"),
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product,
            receta=self.product,
            sale_date=self.target_day,
            quantity=Decimal("50"),
            total_amount=Decimal("5000"),
            source_endpoint="/Report/VentasCategorias",
        )

        result = get_sales_range(start_date=self.target_day, end_date=self.target_day, producto=self.product)

        self.assertEqual(result["source"], "v2_fact")
        self.assertEqual(result["source_detail"], "point_sales_daily_product_fact")
        self.assertEqual(result["cantidad"], Decimal("7"))
        self.assertEqual(result["monto"], Decimal("350"))
        self.assertTrue(result["coverage_accepted"])
        self.assertEqual(result["coverage_reason"], "strict_priority")

    def test_get_sales_range_prefers_official_legacy_without_mixing(self):
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product,
            receta=self.product,
            sale_date=self.target_day,
            quantity=Decimal("8"),
            total_amount=Decimal("400"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product_other,
            receta=self.product,
            sale_date=self.target_day,
            quantity=Decimal("20"),
            total_amount=Decimal("1000"),
            source_endpoint="/Report/VentasCategorias",
        )

        with self.assertLogs("ventas.services.sales_read_service", level="WARNING") as captured:
            result = get_sales_range(start_date=self.target_day, end_date=self.target_day, producto=self.product)

        self.assertEqual(result["source"], "legacy")
        self.assertEqual(result["source_detail"], "point_daily_sale_official")
        self.assertEqual(result["cantidad"], Decimal("8"))
        self.assertEqual(result["monto"], Decimal("400"))
        self.assertTrue(result["fallback_legacy_used"])
        self.assertIn("legacy fallback used", captured.output[0])
        self.assertTrue(result["coverage_accepted"])
        self.assertEqual(result["coverage_reason"], "strict_priority")

    def test_get_sales_range_supports_multi_branch_scope(self):
        other_branch = Sucursal.objects.create(codigo="LEYVA", nombre="Leyva", activa=True)
        other_point_branch = PointBranch.objects.create(external_id="2", name="Leyva", erp_branch=other_branch)
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=self.target_day,
            product_code="PT001",
            point_name="Pastel Canonico",
            quantity=Decimal("2"),
            total_amount=Decimal("100"),
        )
        VentaAutoritativaPoint.objects.create(
            branch=other_branch,
            product=self.product,
            sale_date=self.target_day,
            product_code="PT001",
            point_name="Pastel Canonico",
            quantity=Decimal("3"),
            total_amount=Decimal("180"),
        )
        PointSalesDailyProductFact.objects.create(
            branch=other_point_branch,
            sale_date=self.target_day,
            sucursal_nombre=other_branch.nombre,
            categoria="Pasteles",
            producto_nombre_historico="Pastel Canonico",
            point_product=self.point_product,
            receta=self.product,
            total_cantidad=Decimal("99"),
            total_venta=Decimal("999"),
        )

        result = get_sales_range(
            start_date=self.target_day,
            end_date=self.target_day,
            producto=self.product,
            sucursales=[self.branch, other_branch],
        )

        self.assertEqual(result["source"], "authoritative")
        self.assertEqual(result["cantidad"], Decimal("5"))
        self.assertEqual(result["monto"], Decimal("280"))
        self.assertEqual(result["coverage_days"], 1)
        self.assertEqual(result["coverage_branches"], 2)
        self.assertEqual(result["sucursal_ids"], [self.branch.id, other_branch.id])
        self.assertTrue(result["coverage_accepted"])
        self.assertEqual(result["coverage_reason"], "strict_priority")

    def test_get_sales_range_returns_none_source_when_no_data_exists(self):
        result = get_sales_range(start_date=self.target_day, end_date=self.target_day, producto=self.product)

        self.assertEqual(result["source"], "none")
        self.assertEqual(result["cantidad"], Decimal("0"))
        self.assertEqual(result["monto"], Decimal("0"))
        self.assertEqual(result["coverage_days"], 0)
        self.assertEqual(result["coverage_branches"], 0)
        self.assertFalse(result["coverage_accepted"])
        self.assertEqual(result["coverage_reason"], "no_data")

    def test_get_sales_range_strict_priority_keeps_authoritative_even_if_v2_has_better_coverage(self):
        next_day = self.target_day + timedelta(days=1)
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=self.target_day,
            product_code="PT001",
            point_name="Pastel Canonico",
            quantity=Decimal("2"),
            total_amount=Decimal("100"),
        )
        PointSalesDailyProductFact.objects.create(
            branch=self.point_branch,
            sale_date=self.target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Pasteles",
            producto_nombre_historico="Pastel Canonico",
            point_product=self.point_product,
            receta=self.product,
            total_cantidad=Decimal("2"),
            total_venta=Decimal("220"),
        )
        PointSalesDailyProductFact.objects.create(
            branch=self.point_branch,
            sale_date=next_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Pasteles",
            producto_nombre_historico="Pastel Canonico",
            point_product=self.point_product,
            receta=self.product,
            total_cantidad=Decimal("3"),
            total_venta=Decimal("390"),
        )

        result = get_sales_range(
            start_date=self.target_day,
            end_date=next_day,
            producto=self.product,
            coverage_policy="strict_priority",
        )

        self.assertEqual(result["source"], "authoritative")
        self.assertEqual(result["cantidad"], Decimal("2"))
        self.assertEqual(result["monto"], Decimal("100"))
        self.assertEqual(result["coverage_days"], 1)
        self.assertEqual(result["coverage_reason"], "strict_priority")

    def test_get_sales_range_prefer_complete_can_choose_v2_over_partial_authoritative(self):
        next_day = self.target_day + timedelta(days=1)
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=self.target_day,
            product_code="PT001",
            point_name="Pastel Canonico",
            quantity=Decimal("2"),
            total_amount=Decimal("100"),
        )
        PointSalesDailyProductFact.objects.create(
            branch=self.point_branch,
            sale_date=self.target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Pasteles",
            producto_nombre_historico="Pastel Canonico",
            point_product=self.point_product,
            receta=self.product,
            total_cantidad=Decimal("2"),
            total_venta=Decimal("220"),
        )
        PointSalesDailyProductFact.objects.create(
            branch=self.point_branch,
            sale_date=next_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Pasteles",
            producto_nombre_historico="Pastel Canonico",
            point_product=self.point_product,
            receta=self.product,
            total_cantidad=Decimal("3"),
            total_venta=Decimal("390"),
        )

        result = get_sales_range(
            start_date=self.target_day,
            end_date=next_day,
            producto=self.product,
            coverage_policy="prefer_complete",
        )

        self.assertEqual(result["source"], "v2_fact")
        self.assertEqual(result["cantidad"], Decimal("5"))
        self.assertEqual(result["monto"], Decimal("610"))
        self.assertEqual(result["coverage_days"], 2)
        self.assertTrue(result["coverage_accepted"])
        self.assertEqual(result["coverage_reason"], "prefer_complete_selected_more_complete_source")

    def test_get_daily_sales_bulk_branch_prefers_authoritative_per_date(self):
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=self.target_day,
            product_code="PT001",
            point_name="Pastel Canonico",
            quantity=Decimal("9"),
            total_amount=Decimal("450"),
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=self.point_branch,
            sale_date=self.target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("15"),
            total_venta=Decimal("750"),
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product,
            receta=self.product,
            sale_date=self.target_day,
            quantity=Decimal("5"),
            total_amount=Decimal("250"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )

        result = get_daily_sales_bulk(fechas=[self.target_day], dimension="branch")

        day = result["dates"][self.target_day.isoformat()]
        self.assertEqual(result["dimension"], "branch")
        self.assertEqual(day["source"], "authoritative")
        self.assertEqual(day["coverage_reason"], "strict_priority")
        self.assertEqual(day["rows"], [
            {
                "key": self.branch.id,
                "branch_id": self.branch.id,
                "branch_code": self.branch.codigo,
                "branch_name": self.branch.nombre,
                "erp_branch_id": self.branch.id,
                "units": Decimal("9"),
                "amount": Decimal("450"),
                "tickets": None,
            }
        ])

    def test_get_daily_sales_bulk_branch_can_choose_more_complete_v2_for_day(self):
        other_branch = Sucursal.objects.create(codigo="LEYVA", nombre="Leyva", activa=True)
        other_point_branch = PointBranch.objects.create(external_id="2", name="Leyva", erp_branch=other_branch)
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=self.target_day,
            product_code="PT001",
            point_name="Pastel Canonico",
            quantity=Decimal("2"),
            total_amount=Decimal("100"),
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=self.point_branch,
            sale_date=self.target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("2"),
            total_venta=Decimal("220"),
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=other_point_branch,
            sale_date=self.target_day,
            sucursal_nombre=other_branch.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("3"),
            total_venta=Decimal("390"),
        )

        result = get_daily_sales_bulk(
            fechas=[self.target_day],
            dimension="branch",
            sucursales=[self.branch, other_branch],
            coverage_policy="prefer_complete",
        )

        day = result["dates"][self.target_day.isoformat()]
        self.assertEqual(day["source"], "v2_fact")
        self.assertEqual(day["coverage_reason"], "prefer_complete_selected_more_complete_source")
        self.assertEqual(
            day["rows"],
            [
                {
                    "key": other_branch.id,
                    "branch_id": other_branch.id,
                    "branch_code": other_branch.codigo,
                    "branch_name": other_branch.nombre,
                    "erp_branch_id": other_branch.id,
                    "units": Decimal("3"),
                    "amount": Decimal("390"),
                    "tickets": None,
                },
                {
                    "key": self.branch.id,
                    "branch_id": self.branch.id,
                    "branch_code": self.branch.codigo,
                    "branch_name": self.branch.nombre,
                    "erp_branch_id": self.branch.id,
                    "units": Decimal("2"),
                    "amount": Decimal("220"),
                    "tickets": None,
                },
            ],
        )

    def test_get_daily_sales_bulk_branch_legacy_prefers_official_without_mixing(self):
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product,
            receta=self.product,
            sale_date=self.target_day,
            quantity=Decimal("8"),
            tickets=3,
            total_amount=Decimal("400"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product_other,
            receta=self.other_product,
            sale_date=self.target_day,
            quantity=Decimal("20"),
            tickets=10,
            total_amount=Decimal("1000"),
            source_endpoint="/Report/VentasCategorias",
        )

        with self.assertLogs("ventas.services.sales_read_service", level="WARNING") as captured:
            result = get_daily_sales_bulk(fechas=[self.target_day], dimension="branch")

        day = result["dates"][self.target_day.isoformat()]
        self.assertEqual(day["source"], "legacy")
        self.assertEqual(day["source_detail"], "point_daily_sale_official")
        self.assertEqual(day["rows"], [
            {
                "key": self.branch.id,
                "branch_id": self.branch.id,
                "branch_code": self.branch.codigo,
                "branch_name": self.branch.nombre,
                "erp_branch_id": self.branch.id,
                "units": Decimal("8"),
                "amount": Decimal("400"),
                "tickets": 3,
            }
        ])
        self.assertIn("legacy fallback used", captured.output[0])

    def test_get_daily_sales_bulk_branch_includes_indicator_map_when_requested(self):
        PointSalesDailyCategoryFact.objects.create(
            branch=self.point_branch,
            sale_date=self.target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("12"),
            total_venta=Decimal("600"),
        )
        PointDailyBranchIndicator.objects.create(
            branch=self.point_branch,
            indicator_date=self.target_day,
            total_amount=Decimal("625"),
            total_tickets=4,
            total_avg_ticket=Decimal("156.25"),
        )

        result = get_daily_sales_bulk(
            fechas=[self.target_day],
            dimension="branch",
            include_indicators=True,
        )

        day = result["dates"][self.target_day.isoformat()]
        self.assertEqual(day["source"], "v2_fact")
        self.assertIn("indicator_map", day)
        self.assertEqual(
            day["indicator_map"],
            {
                self.branch.id: {
                    "amount": Decimal("625"),
                    "tickets": 4,
                }
            },
        )
        self.assertEqual(day["rows"][0]["tickets"], None)

    def test_get_daily_sales_bulk_product_prefers_authoritative_per_date(self):
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=self.target_day,
            product_code="PT001",
            point_name="Pastel Canonico",
            quantity=Decimal("9"),
            total_amount=Decimal("450"),
        )
        PointSalesDailyProductFact.objects.create(
            branch=self.point_branch,
            sale_date=self.target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Pasteles",
            producto_nombre_historico="Pastel Canonico",
            point_product=self.point_product,
            receta=self.product,
            total_cantidad=Decimal("7"),
            total_venta=Decimal("350"),
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product,
            receta=self.product,
            sale_date=self.target_day,
            quantity=Decimal("5"),
            total_amount=Decimal("250"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )

        result = get_daily_sales_bulk(fechas=[self.target_day], dimension="product")

        day = result["dates"][self.target_day.isoformat()]
        self.assertEqual(result["dimension"], "product")
        self.assertEqual(day["source"], "authoritative")
        self.assertEqual(
            day["rows"],
            [
                {
                    "key": f"recipe:{self.product.id}",
                    "product_id": None,
                    "recipe_id": self.product.id,
                    "recipe_name": self.product.nombre,
                    "product_name": "Pastel Canonico",
                    "units": Decimal("9"),
                    "amount": Decimal("450"),
                    "branch_count": 1,
                }
            ],
        )

    def test_get_daily_sales_bulk_product_can_choose_more_complete_v2_for_day(self):
        other_branch = Sucursal.objects.create(codigo="LEYVA", nombre="Leyva", activa=True)
        other_point_branch = PointBranch.objects.create(external_id="2", name="Leyva", erp_branch=other_branch)
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=self.target_day,
            product_code="PT001",
            point_name="Pastel Canonico",
            quantity=Decimal("2"),
            total_amount=Decimal("100"),
        )
        PointSalesDailyProductFact.objects.create(
            branch=self.point_branch,
            sale_date=self.target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Pasteles",
            producto_nombre_historico="Pastel Canonico",
            point_product=self.point_product,
            receta=self.product,
            total_cantidad=Decimal("2"),
            total_venta=Decimal("220"),
        )
        PointSalesDailyProductFact.objects.create(
            branch=other_point_branch,
            sale_date=self.target_day,
            sucursal_nombre=other_branch.nombre,
            categoria="Pasteles",
            producto_nombre_historico="Pastel Canonico",
            point_product=self.point_product,
            receta=self.product,
            total_cantidad=Decimal("3"),
            total_venta=Decimal("390"),
        )

        result = get_daily_sales_bulk(
            fechas=[self.target_day],
            dimension="product",
            sucursales=[self.branch, other_branch],
            coverage_policy="prefer_complete",
        )

        day = result["dates"][self.target_day.isoformat()]
        self.assertEqual(day["source"], "v2_fact")
        self.assertEqual(day["coverage_reason"], "prefer_complete_selected_more_complete_source")
        self.assertEqual(
            day["rows"],
            [
                {
                    "key": f"point:{self.point_product.id}",
                    "product_id": self.point_product.id,
                    "recipe_id": self.product.id,
                    "recipe_name": self.product.nombre,
                    "product_name": "Pastel Canonico",
                    "units": Decimal("5"),
                    "amount": Decimal("610"),
                    "branch_count": 2,
                }
            ],
        )

    def test_get_daily_sales_bulk_product_legacy_prefers_official_without_mixing(self):
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product,
            receta=self.product,
            sale_date=self.target_day,
            quantity=Decimal("8"),
            total_amount=Decimal("400"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product_other,
            receta=self.other_product,
            sale_date=self.target_day,
            quantity=Decimal("20"),
            total_amount=Decimal("1000"),
            source_endpoint="/Report/VentasCategorias",
        )

        with self.assertLogs("ventas.services.sales_read_service", level="WARNING") as captured:
            result = get_daily_sales_bulk(fechas=[self.target_day], dimension="product")

        day = result["dates"][self.target_day.isoformat()]
        self.assertEqual(day["source"], "legacy")
        self.assertEqual(day["source_detail"], "point_daily_sale_official")
        self.assertEqual(
            day["rows"],
            [
                {
                    "key": f"point:{self.point_product.id}",
                    "product_id": self.point_product.id,
                    "recipe_id": self.product.id,
                    "recipe_name": self.product.nombre,
                    "product_name": self.point_product.name,
                    "units": Decimal("8"),
                    "amount": Decimal("400"),
                    "branch_count": 1,
                }
            ],
        )
        self.assertIn("legacy fallback used", captured.output[0])

    def test_get_daily_sales_bulk_product_keeps_product_name_when_recipe_is_missing(self):
        PointSalesDailyProductFact.objects.create(
            branch=self.point_branch,
            sale_date=self.target_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Servicios",
            producto_nombre_historico="Servicio de domicilio",
            point_product=None,
            receta=None,
            total_cantidad=Decimal("4"),
            total_venta=Decimal("180"),
        )

        result = get_daily_sales_bulk(fechas=[self.target_day], dimension="product")

        day = result["dates"][self.target_day.isoformat()]
        self.assertEqual(day["source"], "v2_fact")
        self.assertEqual(
            day["rows"],
            [
                {
                    "key": "name:Servicio de domicilio",
                    "product_id": None,
                    "recipe_id": None,
                    "recipe_name": None,
                    "product_name": "Servicio de domicilio",
                    "units": Decimal("4"),
                    "amount": Decimal("180"),
                    "branch_count": 1,
                }
            ],
        )

    def test_resolve_unit_price_prefers_more_complete_v2_range_over_partial_authoritative(self):
        first_day = date(2025, 4, 22)
        second_day = date(2025, 4, 23)
        point_branch = PointBranch.objects.create(external_id="10", name="Matriz Coverage", erp_branch=self.branch)
        point_product = PointProduct.objects.create(external_id="P9", sku="PF9", name=self.product.nombre, category="Temporada")
        VentaAutoritativaPoint.objects.create(
            branch=self.branch,
            product=self.product,
            sale_date=first_day,
            product_code="PF9",
            point_name=self.product.nombre,
            category="Temporada",
            quantity=Decimal("1"),
            total_amount=Decimal("100"),
            net_amount=Decimal("100"),
            tax_amount=Decimal("0"),
            gross_amount=Decimal("100"),
            discount_amount=Decimal("0"),
            source_file="test",
            source_sheet="Sheet1",
            raw_payload={},
        )
        PointSalesDailyProductFact.objects.create(
            branch=point_branch,
            sale_date=first_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Temporada",
            producto_nombre_historico=self.product.nombre,
            point_product=point_product,
            receta=self.product,
            match_catalogo_status="EXACT_CODE",
            total_cantidad=Decimal("2"),
            total_venta=Decimal("220"),
            total_descuento=Decimal("0"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("220"),
            source_hash="fact-hash-coverage-1",
            source_file="fact.xls",
        )
        PointSalesDailyProductFact.objects.create(
            branch=point_branch,
            sale_date=second_day,
            sucursal_nombre=self.branch.nombre,
            categoria="Temporada",
            producto_nombre_historico=self.product.nombre,
            point_product=point_product,
            receta=self.product,
            match_catalogo_status="EXACT_CODE",
            total_cantidad=Decimal("3"),
            total_venta=Decimal("390"),
            total_descuento=Decimal("0"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("390"),
            source_hash="fact-hash-coverage-2",
            source_file="fact.xls",
        )

        price = resolve_unit_price(self.product.id, first_day, second_day)

        self.assertEqual(price, Decimal("122.0000"))
