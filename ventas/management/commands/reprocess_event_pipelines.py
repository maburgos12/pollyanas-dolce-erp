from __future__ import annotations

from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ventas.services.pipeline_audit import (
    collect_event_pipeline_snapshot,
    identify_priority_events,
    write_event_pipeline_audit_report,
)
from recetas.utils.commercial_composition import (
    RULE_BLOQUEADO_POR_AMBIGUEDAD,
    RULE_COMPLEMENTO_OBLIGATORIO,
    RULE_HISTORICO_LEGADO,
    RULE_PRODUCTO_BASE_DIRECTO,
    RULE_SIN_RELACION,
    iter_commercial_validation_rows,
)


class Command(BaseCommand):
    help = (
        "Identifica eventos relevantes, reprocesa el pipeline comercial/operativo y genera "
        "un reporte JSON de auditoría por evento."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--event-id",
            action="append",
            dest="event_ids",
            type=int,
            help="ID de evento específico. Repetible.",
        )
        parser.add_argument(
            "--lookback-days",
            type=int,
            default=60,
            help="Ventana de revisión para eventos sin IDs explícitos.",
        )
        parser.add_argument(
            "--skip-purchases",
            action="store_true",
            help="No regenerar requerimientos de compra.",
        )
        parser.add_argument(
            "--skip-postmortem",
            action="store_true",
            help="No recalcular postmortem aunque aplique.",
        )
        parser.add_argument(
            "--reuse-current-state",
            action="store_true",
            help="No reprocesar pipelines; solo auditar y regenerar el consolidado con el estado actual.",
        )

    def handle(self, *args, **options):
        event_ids = options.get("event_ids") or []
        lookback_days = int(options.get("lookback_days") or 60)
        skip_purchases = bool(options.get("skip_purchases"))
        skip_postmortem = bool(options.get("skip_postmortem"))
        reuse_current_state = bool(options.get("reuse_current_state"))

        events = identify_priority_events(event_ids=event_ids, lookback_days=lookback_days)
        if not events:
            self.stdout.write(self.style.WARNING("No se encontraron eventos relevantes para reprocesar."))
            return

        self.stdout.write(self.style.MIGRATE_HEADING(f"Eventos priorizados: {len(events)}"))
        for event in events:
            self.stdout.write(f"- {event.id} · {event.code} · {event.name} · {event.status}")

        audit_rows: list[dict[str, object]] = []
        module_rows: list[dict[str, object]] = []
        for event in events:
            selection_reason = (
                "explicit_event_ids"
                if event_ids
                else f"status_or_recent_activity_lookback_{lookback_days}d"
            )
            before = collect_event_pipeline_snapshot(event)
            from ventas.views import _event_module_audit_rows, _reprocess_event_for_audit

            if reuse_current_state:
                execution = {
                    "forecast": {"reused": True},
                    "production": {"reused": True},
                    "inputs": {"reused": True},
                    "purchases": {"reused": True},
                    "financials": {"reused": True},
                    "postmortem": {"reused": True},
                    "artifacts": [],
                }
                steps = [{"step": "reuse_current_state", "result": {"enabled": True}}]
            else:
                execution = _reprocess_event_for_audit(
                    event,
                    event.created_by,
                    skip_purchases=skip_purchases,
                    skip_postmortem=skip_postmortem,
                )
                steps = [
                    {"step": "forecast", "result": execution["forecast"]},
                    {"step": "production", "result": execution["production"]},
                    {"step": "inputs", "result": execution["inputs"]},
                    {"step": "purchases", "result": execution["purchases"]},
                    {"step": "financials", "result": execution["financials"]},
                    {"step": "postmortem", "result": execution["postmortem"]},
                    {"step": "artifacts", "result": {"count": len(execution["artifacts"])}},
                ]

            event.refresh_from_db()
            after = collect_event_pipeline_snapshot(event)
            audit_path = write_event_pipeline_audit_report(
                event,
                before=before,
                after=after,
                steps=steps,
                selection_reason=selection_reason,
            )
            semaphores = _event_module_audit_rows(event)
            red_modules = [row["module"] for row in semaphores if row["status"] == "ROJO"]
            yellow_modules = [row["module"] for row in semaphores if row["status"] == "AMARILLO"]
            green_modules = [row["module"] for row in semaphores if row["status"] == "VERDE"]
            audit_rows.append(
                {
                    "event_id": event.id,
                    "code": event.code,
                    "name": event.name,
                    "status": event.status,
                    "forecast_rows": after["forecast"]["rows"],
                    "forecast_total_qty": after["forecast"]["total_qty"],
                    "price_qty_pct": after["coverage"]["price_qty_pct"],
                    "cost_qty_pct": after["coverage"]["cost_qty_pct"],
                    "audit_path": str(audit_path),
                    "red_modules": len(red_modules),
                    "yellow_modules": len(yellow_modules),
                    "green_modules": len(green_modules),
                }
            )
            for row in semaphores:
                module_rows.append(
                    {
                        "event_code": event.code,
                        "event_name": event.name,
                        "module": row["module"],
                        "status": row["status"],
                        "detail": row["detail"],
                    }
                )
            self.stdout.write(
                self.style.SUCCESS(
                    f"Evento {event.id} reprocesado. Forecast {after['forecast']['rows']} filas, "
                    f"artifact audit {audit_path}."
                )
            )
        workbook_path = self._write_audit_workbook(
            event_rows=audit_rows,
            module_rows=module_rows,
            classification_rows=self._classification_rows(),
        )
        self.stdout.write(self.style.SUCCESS(f"Workbook consolidado de auditoría: {workbook_path}"))

    def _classification_rows(self) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for interpretation in iter_commercial_validation_rows():
            rows.append(
                {
                    "sku_actual": interpretation.sku_actual,
                    "producto_actual": interpretation.producto_actual,
                    "clasificacion": interpretation.clasificacion,
                    "sku_base": interpretation.sku_base,
                    "producto_base": interpretation.producto_base,
                    "sku_historico": interpretation.sku_historico,
                    "producto_historico": interpretation.producto_historico,
                    "nota_negocio": interpretation.nota_negocio,
                }
            )
        return rows

    def _write_audit_workbook(
        self,
        *,
        event_rows: list[dict[str, object]],
        module_rows: list[dict[str, object]],
        classification_rows: list[dict[str, object]],
    ) -> Path:
        generated_at = timezone.localtime()
        output_dir = Path(settings.BASE_DIR) / "output" / "spreadsheet" / "validacion_negocio"
        output_dir.mkdir(parents=True, exist_ok=True)
        path = output_dir / f"auditoria_eventos_comerciales_{generated_at.date().isoformat()}.xlsx"

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        ws_events = wb.create_sheet("Eventos")
        ws_modules = wb.create_sheet("Semaforo por modulo")
        ws_blocked = wb.create_sheet("SKU bloqueados")
        ws_classification = wb.create_sheet("Clasificacion maestra")

        red_fill = PatternFill("solid", fgColor="C62828")
        yellow_fill = PatternFill("solid", fgColor="F9A825")
        green_fill = PatternFill("solid", fgColor="2E7D32")
        white_font = Font(color="FFFFFF", bold=True)
        blocked_count = sum(1 for row in classification_rows if row["clasificacion"] == RULE_BLOQUEADO_POR_AMBIGUEDAD)
        direct_count = sum(1 for row in classification_rows if row["clasificacion"] == RULE_PRODUCTO_BASE_DIRECTO)
        historical_count = sum(1 for row in classification_rows if row["clasificacion"] == RULE_HISTORICO_LEGADO)
        complement_count = sum(1 for row in classification_rows if row["clasificacion"] == RULE_COMPLEMENTO_OBLIGATORIO)
        sin_relacion_count = sum(1 for row in classification_rows if row["clasificacion"] == RULE_SIN_RELACION)

        ws_summary.append(["Auditoría final eventos comerciales"])
        ws_summary.append(["Generado", generated_at.isoformat()])
        ws_summary.append([])
        ws_summary.append(["Indicador", "Valor"])
        ws_summary.append(["Eventos auditados", len(event_rows)])
        ws_summary.append(["Módulos en verde", sum(1 for row in module_rows if row["status"] == "VERDE")])
        ws_summary.append(["Módulos en amarillo", sum(1 for row in module_rows if row["status"] == "AMARILLO")])
        ws_summary.append(["Módulos en rojo", sum(1 for row in module_rows if row["status"] == "ROJO")])
        ws_summary.append(["SKU histórico legado", historical_count])
        ws_summary.append(["SKU complemento obligatorio", complement_count])
        ws_summary.append(["SKU producto base directo", direct_count])
        ws_summary.append(["SKU sin relación", sin_relacion_count])
        ws_summary.append(["SKU bloqueado por ambigüedad", blocked_count])

        ws_events.append(
            [
                "Evento ID",
                "Código",
                "Nombre",
                "Estado",
                "Forecast filas",
                "Forecast total",
                "Cobertura precio %",
                "Cobertura costo %",
                "Módulos verde",
                "Módulos amarillo",
                "Módulos rojo",
                "Audit JSON",
            ]
        )
        for row in event_rows:
            ws_events.append(
                [
                    row["event_id"],
                    row["code"],
                    row["name"],
                    row["status"],
                    row["forecast_rows"],
                    row["forecast_total_qty"],
                    row["price_qty_pct"],
                    row["cost_qty_pct"],
                    row["green_modules"],
                    row["yellow_modules"],
                    row["red_modules"],
                    row["audit_path"],
                ]
            )

        ws_modules.append(["Código evento", "Evento", "Módulo", "Semáforo", "Detalle"])
        for row in module_rows:
            ws_modules.append(
                [
                    row["event_code"],
                    row["event_name"],
                    row["module"],
                    row["status"],
                    row["detail"],
                ]
            )

        ws_blocked.append(
            [
                "SKU actual",
                "Producto actual",
                "Clasificación",
                "SKU base",
                "Producto base",
                "SKU histórico",
                "Producto histórico",
                "Nota negocio",
            ]
        )
        for row in classification_rows:
            if row["clasificacion"] != RULE_BLOQUEADO_POR_AMBIGUEDAD:
                continue
            ws_blocked.append(
                [
                    row["sku_actual"],
                    row["producto_actual"],
                    row["clasificacion"],
                    row["sku_base"],
                    row["producto_base"],
                    row["sku_historico"],
                    row["producto_historico"],
                    row["nota_negocio"],
                ]
            )

        ws_classification.append(
            [
                "SKU actual",
                "Producto actual",
                "Clasificación",
                "SKU base",
                "Producto base",
                "SKU histórico",
                "Producto histórico",
                "Nota negocio",
            ]
        )
        for row in classification_rows:
            ws_classification.append(
                [
                    row["sku_actual"],
                    row["producto_actual"],
                    row["clasificacion"],
                    row["sku_base"],
                    row["producto_base"],
                    row["sku_historico"],
                    row["producto_historico"],
                    row["nota_negocio"],
                ]
            )

        for sheet in (ws_events, ws_modules, ws_blocked, ws_classification):
            for cell in sheet[1]:
                cell.font = Font(bold=True)
        for row in ws_modules.iter_rows(min_row=2):
            cell = row[3]
            if cell.value == "ROJO":
                cell.fill = red_fill
                cell.font = white_font
            elif cell.value == "AMARILLO":
                cell.fill = yellow_fill
                cell.font = Font(bold=True)
            elif cell.value == "VERDE":
                cell.fill = green_fill
                cell.font = white_font

        ws_summary.column_dimensions["A"].width = 28
        ws_summary.column_dimensions["B"].width = 22
        for sheet in (ws_events, ws_modules, ws_blocked, ws_classification):
            for column in sheet.columns:
                column_letter = column[0].column_letter
                max_length = max(len(str(cell.value or "")) for cell in column)
                sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 48)

        wb.save(path)
        return path
