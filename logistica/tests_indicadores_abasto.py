from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import Group, User
from django.test import SimpleTestCase, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from core.models import Sucursal
from logistica.models import (
    ParadaEntregaEvidencia,
    ParadaRuta,
    PuntoLogistico,
    RutaCargaChecklist,
    RutaCargaChecklistLinea,
    RutaEntrega,
)
from logistica.exports_indicadores_abasto import build_indicadores_abasto_xlsx
from logistica.services_indicadores_abasto import build_indicadores_abasto
from pos_bridge.models import PointBranch, PointTransferLine


class IndicadoresAbastoContractTests(SimpleTestCase):
    def test_servicio_analitico_publica_constructor_del_reporte(self):
        try:
            from logistica.services_indicadores_abasto import build_indicadores_abasto
        except ModuleNotFoundError:
            build_indicadores_abasto = None

        self.assertIsNotNone(build_indicadores_abasto)
        self.assertTrue(callable(build_indicadores_abasto))


class IndicadoresAbastoServiceTests(TestCase):
    FECHA = date(2026, 7, 21)

    def setUp(self):
        self.sucursal = Sucursal.objects.create(nombre="Plaza Nío", codigo="PNI")
        self.cedis = PointBranch.objects.create(external_id="CEDIS-IND", name="CEDIS")
        self.destino = PointBranch.objects.create(
            external_id="PNI-IND",
            name="Plaza Nío",
            erp_branch=self.sucursal,
        )

    def _transferencia(
        self,
        *,
        suffix,
        solicitado,
        enviado,
        recibido="0",
        enviada=True,
        recibida=False,
        abierta=False,
        actual=True,
        insumo=False,
        unidad="PZA",
        item="Pastel de prueba",
    ):
        momento = timezone.make_aware(datetime(2026, 7, 21, 7, 0))
        return PointTransferLine.objects.create(
            origin_branch=self.cedis,
            destination_branch=self.destino,
            erp_destination_branch=self.sucursal,
            transfer_external_id=f"TR-{suffix}",
            detail_external_id=f"DET-{suffix}",
            source_hash=f"indicadores-{suffix}",
            registered_at=momento,
            sent_at=momento if enviada else None,
            received_at=momento if recibida else None,
            item_code=f"P-{suffix}",
            item_name=item,
            unit=unidad,
            requested_quantity=Decimal(solicitado),
            sent_quantity=Decimal(enviado),
            received_quantity=Decimal(recibido),
            is_insumo=insumo,
            is_received=recibida,
            is_finalized=not abierta,
            is_open=abierta,
            is_current_snapshot=actual,
            raw_payload={"transfer": {"isEnviado": enviada, "isRecibido": recibida}},
        )

    def _vincular_ruta(self, transferencia, *, cargado, recibido):
        punto = PuntoLogistico.objects.create(
            nombre=self.sucursal.nombre,
            tipo=PuntoLogistico.TIPO_SUCURSAL,
            sucursal=self.sucursal,
            latitud="25.570000",
            longitud="-108.470000",
        )
        ruta = RutaEntrega.objects.create(
            nombre="CEDIS - Plaza Nío",
            folio="RUT-IND-001",
            chofer="Luis Repartidor",
            fecha_ruta=self.FECHA,
            estatus=RutaEntrega.ESTATUS_COMPLETADA,
        )
        parada = ParadaRuta.objects.create(ruta=ruta, punto=punto, orden=1)
        checklist = RutaCargaChecklist.objects.create(ruta=ruta)
        linea = RutaCargaChecklistLinea.objects.create(
            checklist=checklist,
            parada=parada,
            point_transfer_line=transferencia,
            transfer_external_id=transferencia.transfer_external_id,
            detail_external_id=transferencia.detail_external_id,
            source_hash=transferencia.source_hash,
            item_code=transferencia.item_code,
            item_name=transferencia.item_name,
            unit=transferencia.unit,
            erp_destination_branch=self.sucursal,
            cantidad_solicitada=transferencia.requested_quantity,
            cantidad_enviada_esperada=transferencia.sent_quantity,
            cantidad_cargada=Decimal(cargado),
            estatus=RutaCargaChecklistLinea.ESTATUS_CARGADA,
        )
        ParadaEntregaEvidencia.objects.create(
            ruta=ruta,
            parada=parada,
            linea_carga=linea,
            tipo=ParadaEntregaEvidencia.TIPO_CONFIRMACION,
            cantidad_entregada=Decimal(recibido),
            client_event_id=f"point-recepcion-{transferencia.source_hash}",
            metadata={"origen": "point_transfer"},
        )
        return ruta

    def test_consolida_sin_compensar_sobrantes_y_separa_pendientes(self):
        mixta = self._transferencia(
            suffix="MIXTA",
            solicitado="10",
            enviado="8",
            recibido="7",
            recibida=True,
        )
        self._vincular_ruta(mixta, cargado="8", recibido="7")
        self._transferencia(
            suffix="CERO",
            solicitado="5",
            enviado="0",
            enviada=True,
            recibida=False,
        )
        self._transferencia(
            suffix="PENDIENTE",
            solicitado="4",
            enviado="4",
            enviada=True,
            recibida=False,
            abierta=True,
        )
        self._transferencia(
            suffix="VIEJA",
            solicitado="99",
            enviado="99",
            recibido="99",
            recibida=True,
            actual=False,
        )
        self._transferencia(
            suffix="INSUMO",
            solicitado="3",
            enviado="3",
            recibido="3",
            recibida=True,
            insumo=True,
            unidad="KG",
        )

        reporte = build_indicadores_abasto(
            fecha_desde=self.FECHA,
            fecha_hasta=self.FECHA,
            tipo="productos",
            unidad="PZA",
        )

        self.assertIn("rows", reporte)
        self.assertIn("totals", reporte)
        self.assertEqual(len(reporte["rows"]), 3)
        self.assertEqual(reporte["totals"]["solicitado"], Decimal("19"))
        self.assertEqual(reporte["totals"]["enviado"], Decimal("12"))
        self.assertEqual(reporte["totals"]["recibido"], Decimal("7"))
        self.assertEqual(reporte["totals"]["pendientes"], 1)
        self.assertEqual(reporte["totals"]["brecha_abasto"], Decimal("7"))
        self.assertEqual(reporte["totals"]["brecha_entrega"], Decimal("1"))
        self.assertEqual(reporte["totals"]["porcentaje_abasto"], Decimal("63.2"))
        self.assertEqual(reporte["totals"]["porcentaje_entrega"], Decimal("87.5"))
        self.assertEqual(reporte["totals"]["porcentaje_total_evaluado"], Decimal("46.7"))
        estados = {row["transfer_external_id"]: row["estado"] for row in reporte["rows"]}
        self.assertEqual(estados["TR-MIXTA"], "BRECHA_MIXTA")
        self.assertEqual(estados["TR-CERO"], "NO_SURTIDO")
        self.assertEqual(estados["TR-PENDIENTE"], "PENDIENTE")
        fila_mixta = next(row for row in reporte["rows"] if row["transfer_external_id"] == "TR-MIXTA")
        self.assertEqual(fila_mixta["cargado"], Decimal("8"))
        self.assertEqual(fila_mixta["ruta_folio"], "RUT-IND-001")
        self.assertEqual(reporte["unidades"], ["PZA"])
        self.assertFalse(reporte["mezcla_unidades"])
        self.assertEqual(reporte["por_sucursal"][0]["etiqueta"], "Plaza Nío")
        self.assertEqual(reporte["por_dia"][0]["clave"], (self.FECHA, "PZA"))
        self.assertEqual(reporte["por_producto"][0]["unidad"], "PZA")
        self.assertEqual(reporte["por_ruta"][0]["etiqueta"], "RUT-IND-001 · CEDIS - Plaza Nío")
        self.assertEqual(reporte["por_ruta"][0]["responsable"], "Luis Repartidor")
        self.assertEqual(reporte["por_ruta"][0]["cargado"], Decimal("8"))
        workbook = load_workbook(
            BytesIO(
                build_indicadores_abasto_xlsx(
                    reporte,
                    {"fecha_desde": self.FECHA, "fecha_hasta": self.FECHA, "unidad": "PZA"},
                )
            ),
            read_only=True,
        )
        self.assertEqual(workbook["Por producto"]["B2"].value, "Pastel de prueba")

    def test_reporta_unidades_por_separado_cuando_no_hay_filtro(self):
        self._transferencia(
            suffix="PIEZAS", solicitado="2", enviado="2", recibido="2", recibida=True
        )
        self._transferencia(
            suffix="KILOS",
            solicitado="3",
            enviado="3",
            recibido="3",
            recibida=True,
            insumo=True,
            unidad="KG",
        )

        reporte = build_indicadores_abasto(
            fecha_desde=self.FECHA,
            fecha_hasta=self.FECHA,
            tipo="todos",
        )

        self.assertEqual(reporte["unidades"], ["KG", "PZA"])
        self.assertTrue(reporte["mezcla_unidades"])
        self.assertEqual([grupo["unidad"] for grupo in reporte["por_unidad"]], ["KG", "PZA"])

    def test_excel_conserva_codigos_y_neutraliza_formulas_en_texto(self):
        self._transferencia(
            suffix="00123",
            solicitado="1",
            enviado="1",
            recibido="1",
            recibida=True,
            item="=HYPERLINK(\"https://example.invalid\")",
        )
        reporte = build_indicadores_abasto(
            fecha_desde=self.FECHA,
            fecha_hasta=self.FECHA,
            tipo="productos",
            unidad="PZA",
        )

        workbook = load_workbook(
            BytesIO(
                build_indicadores_abasto_xlsx(
                    reporte,
                    {"fecha_desde": self.FECHA, "fecha_hasta": self.FECHA, "unidad": "PZA"},
                )
            ),
            read_only=True,
            data_only=False,
        )

        detail = workbook["Detalle"]
        self.assertEqual(detail["C2"].value, "P-00123")
        self.assertEqual(detail["D2"].value, "'=HYPERLINK(\"https://example.invalid\")")


class IndicadoresAbastoViewsTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="indicadores.logistica", password="pass123")
        group, _ = Group.objects.get_or_create(name="LOGISTICA")
        self.user.groups.add(group)
        self.client.force_login(self.user)

    def test_pantalla_publica_analisis_y_descarga_para_logistica(self):
        response = self.client.get("/logistica/indicadores-abasto/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Indicadores de abasto")
        self.assertContains(response, "Solicitado")
        self.assertContains(response, "Enviado")
        self.assertContains(response, "Recibido")
        self.assertContains(response, "Descargar Excel")
        self.assertContains(response, "Dónde se queda el producto")

    def test_url_directa_rechaza_usuario_sin_acceso_a_logistica(self):
        outsider = User.objects.create_user(username="sin.logistica", password="pass123")
        self.client.force_login(outsider)

        response = self.client.get("/logistica/indicadores-abasto/")

        self.assertEqual(response.status_code, 403)

    def test_excel_entrega_cinco_hojas_analiticas(self):
        response = self.client.get("/logistica/indicadores-abasto/exportar/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ["Resumen", "Por sucursal", "Por día", "Por producto", "Detalle"],
        )
        headers = [cell.value for cell in next(workbook["Detalle"].iter_rows())]
        self.assertIn("Cargado", headers)
        self.assertIn("Estado / causa", headers)

    def test_navegacion_horizontal_incluye_indicadores(self):
        response = self.client.get(reverse("logistica:home"))

        self.assertContains(response, reverse("logistica:indicadores_abasto"))
        self.assertContains(response, "Indicadores de abasto")
