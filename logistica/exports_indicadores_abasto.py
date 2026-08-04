from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


VINO = "8B2252"
DORADO = "C9A84C"
BLANCO = "FFFAF5"


def _value(value):
    if value is None:
        return "N/A"
    if isinstance(value, (tuple, list)):
        return " · ".join(str(part) for part in value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, (Decimal, int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _sheet(workbook, title, headers, rows):
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append([_value(value) for value in row])
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=VINO)
        cell.font = Font(color=BLANCO, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        values = [len(str(cell.value or "")) for cell in column]
        sheet.column_dimensions[column[0].column_letter].width = min(max(values + [10]) + 2, 42)
    return sheet


def build_indicadores_abasto_xlsx(report, filters):
    workbook = Workbook()
    workbook.remove(workbook.active)
    totals = report["totals"]
    summary_rows = [
        ["Periodo", f'{filters["fecha_desde"]} al {filters["fecha_hasta"]}', ""],
    ]
    if report["mezcla_unidades"]:
        for unit in report["por_unidad"]:
            summary_rows.extend(
                [
                    [f'Solicitado evaluado · {unit["unidad"]}', unit["solicitado_evaluado"], unit["unidad"]],
                    [f'Enviado evaluado · {unit["unidad"]}', unit["enviado_evaluado"], unit["unidad"]],
                    [f'Recibido evaluado · {unit["unidad"]}', unit["recibido_evaluado"], unit["unidad"]],
                    [f'Cumplimiento total · {unit["unidad"]}', unit["porcentaje_total_evaluado"], "% evaluado"],
                    [f'Pendiente de envío CEDIS · {unit["unidad"]}', unit["pendientes_envio_lineas"], "líneas abiertas"],
                    [f'Pendiente de recepción sucursal · {unit["unidad"]}', unit["pendientes_recepcion_lineas"], "líneas abiertas"],
                ]
            )
    else:
        summary_rows.extend(
            [
                ["Solicitado evaluado", totals["solicitado_evaluado"], "Solo transferencias cerradas"],
                ["Enviado evaluado", totals["enviado_evaluado"], "Limitado a lo solicitado; sin compensar sobrantes"],
                ["Recibido evaluado", totals["recibido_evaluado"], "Limitado a lo enviado y solicitado"],
                ["Cumplimiento de envío CEDIS", totals["porcentaje_abasto"], "Enviado evaluado / solicitado evaluado"],
                ["Cumplimiento de recepción sucursal", totals["porcentaje_entrega"], "Recibido evaluado / enviado evaluado"],
                ["Cumplimiento final solicitado–recibido", totals["porcentaje_total_evaluado"], "Recibido evaluado / solicitado evaluado"],
                ["Brecha confirmada CEDIS", totals["brecha_abasto"], "Solicitado evaluado - enviado evaluado"],
                ["Brecha confirmada de entrega", totals["brecha_entrega"], "Enviado evaluado - recibido evaluado"],
                ["Pendiente de envío CEDIS", totals["pendientes_envio_lineas"], f'{totals["pendientes_envio_solicitado"]} solicitado en líneas abiertas'],
                ["Pendiente de recepción sucursal", totals["pendientes_recepcion_lineas"], f'{totals["pendientes_recepcion_enviado"]} enviado en líneas abiertas'],
                ["Sobrante enviado", totals["sobrante_envio"], "Se reporta aparte; no compensa faltantes"],
                ["Sobrante recibido", totals["sobrante_recepcion"], "Se reporta aparte; no compensa faltantes"],
            ]
        )
    summary_rows.append(["Líneas abiertas fuera del cálculo", totals["pendientes"], "No se cuentan como incumplidas ni como brecha confirmada"])
    _sheet(
        workbook,
        "Resumen",
        ["Indicador", "Valor", "Unidad / criterio"],
        summary_rows,
    )
    group_headers = [
        "Clave", "Nombre", "Unidad", "Solicitado evaluado", "Enviado evaluado", "Recibido evaluado",
        "Brecha CEDIS confirmada", "Brecha entrega confirmada", "Brecha de carga", "Brecha de ruta",
        "% envío CEDIS", "% recepción sucursal",
        "% final solicitado-recibido", "Pendientes envío CEDIS", "Pendientes recepción sucursal",
    ]
    def group_values(group):
        return [
            group["clave"], group["etiqueta"], group["unidad"], group["solicitado_evaluado"], group["enviado_evaluado"],
            group["recibido_evaluado"], group["brecha_abasto"], group["brecha_entrega"], group["brecha_carga"],
            group["brecha_ruta"], group["porcentaje_abasto"],
            group["porcentaje_entrega"], group["porcentaje_total_evaluado"], group["pendientes_envio_lineas"],
            group["pendientes_recepcion_lineas"],
        ]
    _sheet(workbook, "Por sucursal", group_headers, [group_values(row) for row in report["por_sucursal"]])
    _sheet(workbook, "Por día", group_headers, [group_values(row) for row in report["por_dia"]])
    _sheet(workbook, "Por producto", group_headers, [group_values(row) for row in report["por_producto"]])
    detail_headers = [
        "Fecha", "Sucursal", "Código", "Producto / insumo", "Unidad", "Solicitado", "Enviado", "Cargado",
        "Recibido", "Brecha CEDIS confirmada", "Brecha entrega confirmada", "Brecha de carga", "Brecha de ruta",
        "Ruta", "Repartidor", "Estado / causa",
        "Responsable del siguiente paso", "Pendiente",
        "Transferencia Point", "Detalle Point", "Enviado por", "Recibido por", "Recibido en",
    ]
    detail_rows = [
        [
            row["fecha"], row["sucursal"], row["item_code"], row["item_name"], row["unidad"],
            row["solicitado"], row["enviado"], row["cargado"], row["recibido"], row["brecha_abasto"],
            row["brecha_entrega"], row["brecha_carga"], row["brecha_ruta"], row["ruta_folio"], row["repartidor"], row["estado"],
            row["responsable_siguiente_paso"], "Sí" if row["pendiente"] else "No", row["transfer_external_id"], row["detail_external_id"],
            row["sent_by"], row["received_by"], row["received_at"],
        ]
        for row in report["rows"]
    ]
    _sheet(workbook, "Detalle", detail_headers, detail_rows)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
