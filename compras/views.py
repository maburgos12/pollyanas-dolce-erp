import csv
import calendar
import re
from collections import defaultdict
from io import BytesIO
from decimal import Decimal, InvalidOperation
from io import StringIO
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db.models import Count, Q, Sum
from django.urls import reverse
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.text import slugify
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook

from core.access import can_manage_compras, can_view_compras
from core.audit import log_event
from inventario.models import ExistenciaInsumo, MovimientoInventario
from maestros.models import CostoInsumo, Insumo, Proveedor
from maestros.utils.canonical_catalog import (
    canonical_insumo,
    canonical_insumo_by_id,
    canonicalized_active_insumos,
    enterprise_readiness_profile,
    latest_costo_canonico,
    usage_maps_for_insumo_ids,
)
from recetas.models import LineaReceta, PlanProduccion, SolicitudVenta, VentaHistorica
from recetas.utils.matching import match_insumo
from recetas.utils.normalizacion import normalizar_nombre

from .models import (
    OrdenCompra,
    PresupuestoCompraCategoria,
    PresupuestoCompraPeriodo,
    PresupuestoCompraProveedor,
    RecepcionCompra,
    SolicitudCompra,
)

IMPORT_PREVIEW_SESSION_KEY = "compras_solicitudes_import_preview"


def _to_decimal(value: str, default: str = "0") -> Decimal:
    try:
        return Decimal(value or default)
    except (InvalidOperation, TypeError):
        return Decimal(default)


def _extract_plan_id_from_scope(value: str | None) -> str:
    text = (value or "").strip()
    if not text:
        return ""
    if text.isdigit():
        return text
    match = re.search(r"PLAN_PRODUCCION:(\d+)", text)
    return match.group(1) if match else ""


def _build_plan_scope_context(
    *,
    source_filter: str = "",
    plan_filter: str = "",
    q_filter: str = "",
    current_view: str = "solicitudes",
    closure_key_filter: str = "all",
    handoff_key_filter: str = "all",
    master_class_filter: str = "all",
    master_missing_filter: str = "all",
    session: dict | None = None,
) -> dict | None:
    resolved_plan_id = ""
    if source_filter == "plan" and plan_filter:
        resolved_plan_id = _extract_plan_id_from_scope(plan_filter)
    if not resolved_plan_id:
        resolved_plan_id = _extract_plan_id_from_scope(q_filter)
    if not resolved_plan_id:
        return None

    plan = PlanProduccion.objects.filter(id=int(resolved_plan_id)).only("id", "nombre", "fecha_produccion").first()
    if not plan:
        return None

    def _plan_query(extra: dict[str, object] | None = None) -> str:
        params: dict[str, object] = {"source": "plan", "plan_id": plan.id}
        if extra:
            params.update({key: value for key, value in extra.items() if value not in (None, "", "all")})
        return urlencode(params)

    solicitudes_query = _plan_query()
    plan_scope = f"PLAN_PRODUCCION:{plan.id}"
    solicitudes_qs = SolicitudCompra.objects.filter(area=plan_scope)
    ordenes_qs = OrdenCompra.objects.filter(
        Q(referencia=plan_scope) | Q(solicitud__area=plan_scope)
    )
    recepciones_qs = RecepcionCompra.objects.filter(
        Q(orden__referencia=plan_scope) | Q(orden__solicitud__area=plan_scope)
    )
    solicitudes_items = list(solicitudes_qs.select_related("insumo", "proveedor_sugerido")[:300])
    open_orders_by_solicitud = {
        orden.solicitud_id: orden
        for orden in OrdenCompra.objects.filter(solicitud_id__in=[s.id for s in solicitudes_items])
        .exclude(estatus=OrdenCompra.STATUS_CERRADA)
        .select_related("solicitud")
        .order_by("-creado_en")
    } if solicitudes_items else {}
    for solicitud in solicitudes_items:
        open_order = open_orders_by_solicitud.get(solicitud.id)
        solicitud.has_open_order = bool(open_order)
        solicitud.open_order_folio = open_order.folio if open_order else ""
        _enrich_solicitud_workflow(solicitud)

    ordenes_items = list(ordenes_qs.select_related("proveedor", "solicitud")[:300])
    recepciones_cerradas_por_orden = {
        row["orden_id"]: row["total"]
        for row in (
            RecepcionCompra.objects.filter(orden_id__in=[o.id for o in ordenes_items], estatus=RecepcionCompra.STATUS_CERRADA)
            .values("orden_id")
            .annotate(total=Count("id"))
        )
    } if ordenes_items else {}
    for orden in ordenes_items:
        _enrich_orden_workflow(orden, int(recepciones_cerradas_por_orden.get(orden.id) or 0))

    recepciones_items = list(recepciones_qs.select_related("orden", "orden__proveedor")[:300])
    for recepcion in recepciones_items:
        _enrich_recepcion_workflow(recepcion)

    solicitudes_pendientes = solicitudes_qs.filter(
        estatus__in=[SolicitudCompra.STATUS_BORRADOR, SolicitudCompra.STATUS_EN_REVISION]
    ).count()
    ordenes_abiertas = ordenes_qs.exclude(estatus=OrdenCompra.STATUS_CERRADA).count()
    recepciones_abiertas = recepciones_qs.exclude(estatus=RecepcionCompra.STATUS_CERRADA).count()
    blocked_total = solicitudes_pendientes + ordenes_abiertas + recepciones_abiertas
    blocked_solicitudes = sum(1 for item in solicitudes_items if getattr(item, "has_workflow_blockers", False))
    blocked_ordenes = sum(1 for item in ordenes_items if getattr(item, "has_workflow_blockers", False))
    blocked_recepciones = sum(1 for item in recepciones_items if getattr(item, "has_workflow_blockers", False))
    approved_ready = sum(
        1
        for item in solicitudes_items
        if item.estatus == SolicitudCompra.STATUS_APROBADA and not getattr(item, "has_open_order", False) and not getattr(item, "has_workflow_blockers", False)
    )
    confirmed_ready = sum(
        1
        for item in ordenes_items
        if item.estatus == OrdenCompra.STATUS_CONFIRMADA and not getattr(item, "has_workflow_blockers", False)
    )
    diferencias_abiertas = sum(1 for item in recepciones_items if item.estatus == RecepcionCompra.STATUS_DIFERENCIAS)
    summary_cards = [
        {
            "label": "Solicitudes",
            "count": solicitudes_qs.count(),
            "open": solicitudes_pendientes,
            "tone": "warning" if solicitudes_pendientes else ("primary" if solicitudes_qs.exists() else "warning"),
        },
        {
            "label": "Órdenes",
            "count": ordenes_qs.count(),
            "open": ordenes_abiertas,
            "tone": "warning" if ordenes_abiertas else ("primary" if ordenes_qs.exists() else "warning"),
        },
        {
            "label": "Recepciones",
            "count": recepciones_qs.count(),
            "open": recepciones_abiertas,
            "tone": "warning" if recepciones_abiertas else ("primary" if recepciones_qs.exists() else "warning"),
        },
    ]
    summary_label = "Con bloqueos" if blocked_total else "Sin bloqueos"
    if current_view == "ordenes":
        priority = [
            "blocked_ordenes",
            "approved_ready",
            "ordenes_abiertas",
            "blocked_solicitudes",
            "solicitudes_pendientes",
            "blocked_recepciones",
            "diferencias_abiertas",
            "recepciones_abiertas",
            "closed",
        ]
    elif current_view == "recepciones":
        priority = [
            "blocked_recepciones",
            "diferencias_abiertas",
            "recepciones_abiertas",
            "blocked_ordenes",
            "confirmed_ready",
            "blocked_solicitudes",
            "solicitudes_pendientes",
            "closed",
        ]
    else:
        priority = [
            "blocked_solicitudes",
            "solicitudes_pendientes",
            "blocked_ordenes",
            "approved_ready",
            "blocked_recepciones",
            "diferencias_abiertas",
            "recepciones_abiertas",
            "ordenes_abiertas",
            "closed",
        ]

    current_stage = priority[-1]
    for key in priority:
        if key == "closed":
            current_stage = "closed"
            break
        value = {
            "blocked_solicitudes": blocked_solicitudes,
            "solicitudes_pendientes": solicitudes_pendientes,
            "blocked_ordenes": blocked_ordenes,
            "approved_ready": approved_ready,
            "ordenes_abiertas": ordenes_abiertas,
            "blocked_recepciones": blocked_recepciones,
            "diferencias_abiertas": diferencias_abiertas,
            "recepciones_abiertas": recepciones_abiertas,
            "confirmed_ready": confirmed_ready,
        }.get(key, 0)
        if value:
            current_stage = key
            break

    if current_stage == "blocked_solicitudes":
        stage_label = "Validación de solicitudes"
        stage_tone = "danger"
        stage_detail = f"{blocked_solicitudes} solicitudes del plan están bloqueadas por maestro ERP o costo."
        next_action = {
            "label": "Corregir solicitudes bloqueadas",
            "detail": "Revisa costo, proveedor y readiness del artículo antes de generar OC.",
            "url": f"{reverse('compras:solicitudes')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': 'BLOCKED_ERP'})}",
        }
    elif current_stage == "solicitudes_pendientes":
        stage_label = "Solicitudes en captura/validación"
        stage_tone = "warning"
        stage_detail = f"Quedan {solicitudes_pendientes} solicitudes sin liberar dentro del plan."
        next_action = {
            "label": "Liberar solicitudes pendientes",
            "detail": "Completa revisión y aprueba solicitudes para pasar a órdenes.",
            "url": f"{reverse('compras:solicitudes')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': SolicitudCompra.STATUS_EN_REVISION})}",
        }
    elif current_stage == "blocked_ordenes":
        stage_label = "Órdenes bloqueadas ERP"
        stage_tone = "danger"
        stage_detail = f"{blocked_ordenes} órdenes del plan no están listas para proveedor o recepción."
        next_action = {
            "label": "Corregir órdenes bloqueadas",
            "detail": "Ajusta proveedor, monto o entrega estimada para continuar el flujo.",
            "url": f"{reverse('compras:ordenes')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': 'BLOCKED_ERP'})}",
        }
    elif current_stage == "approved_ready":
        stage_label = "Listo para emitir órdenes"
        stage_tone = "success"
        stage_detail = f"Hay {approved_ready} solicitudes aprobadas listas para convertirse en órdenes."
        next_action = {
            "label": "Generar órdenes del plan",
            "detail": "Convierte solicitudes aprobadas en órdenes antes de perder ventana de compra.",
            "url": f"{reverse('compras:solicitudes')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': 'APPROVED_READY'})}",
        }
    elif current_stage == "blocked_recepciones":
        stage_label = "Recepciones bloqueadas ERP"
        stage_tone = "danger"
        stage_detail = f"{blocked_recepciones} recepciones no pueden cerrar ni aplicar inventario."
        next_action = {
            "label": "Resolver recepciones bloqueadas",
            "detail": "Completa fecha, observaciones o justificación para aplicar inventario.",
            "url": f"{reverse('compras:recepciones')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': 'BLOCKED_ERP'})}",
        }
    elif current_stage == "diferencias_abiertas":
        stage_label = "Recepciones con diferencias"
        stage_tone = "warning"
        stage_detail = f"Quedan {diferencias_abiertas} recepciones con diferencias por cerrar."
        next_action = {
            "label": "Cerrar recepciones con diferencias",
            "detail": "Resuelve diferencias para no dejar inventario y auditoría abiertos.",
            "url": f"{reverse('compras:recepciones')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': RecepcionCompra.STATUS_DIFERENCIAS})}",
        }
    elif current_stage == "recepciones_abiertas":
        stage_label = "Recepciones en curso"
        stage_tone = "primary"
        stage_detail = f"Hay {recepciones_abiertas} recepciones abiertas dentro del plan."
        next_action = {
            "label": "Completar recepciones abiertas",
            "detail": "Cierra y aplica inventario para terminar el ciclo documental.",
            "url": f"{reverse('compras:recepciones')}?{urlencode({'source': 'plan', 'plan_id': plan.id})}",
        }
    elif current_stage == "ordenes_abiertas":
        stage_label = "Seguimiento a órdenes"
        stage_tone = "primary"
        stage_detail = f"Hay {ordenes_abiertas} órdenes en curso esperando confirmación o recepción."
        next_action = {
            "label": "Dar seguimiento a órdenes",
            "detail": "Revisa órdenes abiertas del plan y avanza a recepción.",
            "url": f"{reverse('compras:ordenes')}?{urlencode({'source': 'plan', 'plan_id': plan.id})}",
        }
    elif current_stage == "confirmed_ready":
        stage_label = "Órdenes listas para recepción"
        stage_tone = "success"
        stage_detail = f"Hay {confirmed_ready} órdenes confirmadas listas para registrar recepción."
        next_action = {
            "label": "Registrar recepciones del plan",
            "detail": "Confirma entrada física y avanza a recepción para cerrar el ciclo.",
            "url": f"{reverse('compras:ordenes')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': OrdenCompra.STATUS_CONFIRMADA})}",
        }
    else:
        stage_label = "Cierre documental"
        stage_tone = "success"
        stage_detail = "El plan no tiene bloqueos ni documentos abiertos en compras."
        next_action = {
            "label": "Volver al plan",
            "detail": "El flujo documental está al día.",
            "url": f"{reverse('recetas:plan_produccion')}?{urlencode({'plan_id': plan.id})}",
        }

    drilldown_actions = []
    if blocked_solicitudes:
        drilldown_actions.append({
            "label": "Solicitudes bloqueadas",
            "count": blocked_solicitudes,
            "tone": "danger",
            "url": f"{reverse('compras:solicitudes')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': 'BLOCKED_ERP'})}",
        })
    if approved_ready:
        drilldown_actions.append({
            "label": "Solicitudes listas para OC",
            "count": approved_ready,
            "tone": "success",
            "url": f"{reverse('compras:solicitudes')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': 'APPROVED_READY'})}",
        })
    if blocked_ordenes:
        drilldown_actions.append({
            "label": "Órdenes bloqueadas",
            "count": blocked_ordenes,
            "tone": "danger",
            "url": f"{reverse('compras:ordenes')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': 'BLOCKED_ERP'})}",
        })
    if confirmed_ready:
        drilldown_actions.append({
            "label": "Órdenes listas para recepción",
            "count": confirmed_ready,
            "tone": "success",
            "url": f"{reverse('compras:ordenes')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': OrdenCompra.STATUS_CONFIRMADA})}",
        })
    if blocked_recepciones:
        drilldown_actions.append({
            "label": "Recepciones bloqueadas",
            "count": blocked_recepciones,
            "tone": "danger",
            "url": f"{reverse('compras:recepciones')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': 'BLOCKED_ERP'})}",
        })
    if diferencias_abiertas:
        drilldown_actions.append({
            "label": "Recepciones con diferencias",
            "count": diferencias_abiertas,
            "tone": "warning",
            "url": f"{reverse('compras:recepciones')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': RecepcionCompra.STATUS_DIFERENCIAS})}",
        })

    pipeline_steps = [
        {
            "label": "Solicitudes",
            "count": solicitudes_qs.count(),
            "open": solicitudes_pendientes,
            "blocked": blocked_solicitudes,
            "closed": max(solicitudes_qs.count() - solicitudes_pendientes - blocked_solicitudes, 0),
            "status_label": "Bloqueadas" if blocked_solicitudes else ("Por atender" if solicitudes_pendientes else ("Liberadas" if solicitudes_qs.exists() else "Sin generar")),
            "tone": "danger" if blocked_solicitudes else ("warning" if solicitudes_pendientes else ("success" if solicitudes_qs.exists() else "muted")),
            "url": f"{reverse('compras:solicitudes')}?{urlencode({'source': 'plan', 'plan_id': plan.id})}",
        },
        {
            "label": "Órdenes",
            "count": ordenes_qs.count(),
            "open": ordenes_abiertas,
            "blocked": blocked_ordenes,
            "closed": max(ordenes_qs.count() - ordenes_abiertas - blocked_ordenes, 0),
            "status_label": "Bloqueadas" if blocked_ordenes else ("Abiertas" if ordenes_abiertas else ("Cerradas" if ordenes_qs.exists() else "Sin generar")),
            "tone": "danger" if blocked_ordenes else ("primary" if ordenes_abiertas else ("success" if ordenes_qs.exists() else "muted")),
            "url": f"{reverse('compras:ordenes')}?{urlencode({'source': 'plan', 'plan_id': plan.id})}",
        },
        {
            "label": "Recepciones",
            "count": recepciones_qs.count(),
            "open": recepciones_abiertas,
            "blocked": blocked_recepciones + diferencias_abiertas,
            "closed": max(recepciones_qs.count() - recepciones_abiertas - blocked_recepciones - diferencias_abiertas, 0),
            "status_label": "Bloqueadas" if blocked_recepciones else ("Diferencias" if diferencias_abiertas else ("Abiertas" if recepciones_abiertas else ("Cerradas" if recepciones_qs.exists() else "Sin generar"))),
            "tone": "danger" if blocked_recepciones else ("warning" if diferencias_abiertas else ("primary" if recepciones_abiertas else ("success" if recepciones_qs.exists() else "muted"))),
            "url": f"{reverse('compras:recepciones')}?{urlencode({'source': 'plan', 'plan_id': plan.id})}",
        },
    ]
    for item in pipeline_steps:
        count = int(item.get("count") or 0)
        blocked = int(item.get("blocked") or 0)
        closed = int(item.get("closed") or 0)
        open_items = int(item.get("open") or 0)
        actionable_total = blocked + open_items + closed
        progress_pct = int(round((closed / actionable_total) * 100)) if actionable_total > 0 else 0
        if blocked > 0:
            semaphore_label = "Rojo"
            semaphore_tone = "danger"
        elif open_items > 0:
            semaphore_label = "Amarillo"
            semaphore_tone = "warning" if item["label"] != "Órdenes" else "primary"
        elif count > 0:
            semaphore_label = "Verde"
            semaphore_tone = "success"
        else:
            semaphore_label = "Sin documentos"
            semaphore_tone = "muted"
        item["progress_pct"] = progress_pct
        item["semaphore_label"] = semaphore_label
        item["semaphore_tone"] = semaphore_tone
        item["is_active"] = item["label"] == (
            "Solicitudes"
            if current_stage in {"blocked_solicitudes", "solicitudes_pendientes", "approved_ready"}
            else "Órdenes"
            if current_stage in {"blocked_ordenes", "ordenes_abiertas", "confirmed_ready"}
            else "Recepciones"
            if current_stage in {"blocked_recepciones", "diferencias_abiertas", "recepciones_abiertas", "closed"}
            else "Solicitudes"
        )
        if item["label"] == "Solicitudes":
            if blocked > 0:
                item["action_label"] = "Resolver bloqueos"
                item["action_detail"] = "Corrige solicitudes bloqueadas para habilitar emisión de órdenes."
                item["action_url"] = f"{reverse('compras:solicitudes')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': 'BLOCKED_ERP'})}"
            elif open_items > 0:
                item["action_label"] = "Liberar pendientes"
                item["action_detail"] = "Termina captura o validación antes de avanzar a órdenes."
                item["action_url"] = f"{reverse('compras:solicitudes')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': SolicitudCompra.STATUS_BORRADOR})}"
            else:
                item["action_label"] = "Abrir solicitudes"
                item["action_detail"] = "La etapa ya está liberada para este plan."
                item["action_url"] = item["url"]
        elif item["label"] == "Órdenes":
            if blocked > 0:
                item["action_label"] = "Corregir órdenes"
                item["action_detail"] = "Completa proveedor, monto y datos documentales."
                item["action_url"] = f"{reverse('compras:ordenes')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': 'BLOCKED_ERP'})}"
            elif open_items > 0:
                item["action_label"] = "Dar seguimiento"
                item["action_detail"] = "Confirma órdenes o prepáralas para recepción."
                item["action_url"] = item["url"]
            else:
                item["action_label"] = "Abrir órdenes"
                item["action_detail"] = "La etapa documental de órdenes ya está cerrada."
                item["action_url"] = item["url"]
        else:
            if blocked > 0:
                item["action_label"] = "Resolver recepciones"
                item["action_detail"] = "Atiende bloqueos ERP o diferencias antes de cerrar el plan."
                item["action_url"] = (
                    f"{reverse('compras:recepciones')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': 'BLOCKED_ERP'})}"
                    if blocked_recepciones
                    else f"{reverse('compras:recepciones')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'estatus': RecepcionCompra.STATUS_DIFERENCIAS})}"
                )
            elif open_items > 0:
                item["action_label"] = "Cerrar recepciones"
                item["action_detail"] = "Aplica inventario y confirma conformidad para terminar."
                item["action_url"] = item["url"]
            else:
                item["action_label"] = "Abrir recepciones"
                item["action_detail"] = "La etapa de recepción ya está cerrada."
                item["action_url"] = item["url"]

    def _stage_focus_row_for_solicitud(item: SolicitudCompra) -> dict:
        return {
            "scope": "Solicitud",
            "folio": item.folio,
            "status": item.get_estatus_display(),
            "detail": f"{item.insumo.nombre if item.insumo_id else 'Sin artículo'} · {item.solicitante}",
            "action_label": "Abrir solicitudes",
            "action_url": f"{reverse('compras:solicitudes')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'q': item.folio})}",
        }

    def _stage_focus_row_for_orden(item: OrdenCompra) -> dict:
        return {
            "scope": "Orden",
            "folio": item.folio,
            "status": item.get_estatus_display(),
            "detail": f"{item.proveedor.nombre if item.proveedor_id else 'Sin proveedor'} · monto {Decimal(str(item.monto_estimado or 0)):.2f}",
            "action_label": "Abrir órdenes",
            "action_url": f"{reverse('compras:ordenes')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'q': item.folio})}",
        }

    def _stage_focus_row_for_recepcion(item: RecepcionCompra) -> dict:
        return {
            "scope": "Recepción",
            "folio": item.folio,
            "status": item.get_estatus_display(),
            "detail": f"{item.orden.folio if item.orden_id else 'Sin orden'} · conformidad {Decimal(str(item.conformidad_pct or 0)):.0f}%",
            "action_label": "Abrir recepciones",
            "action_url": f"{reverse('compras:recepciones')}?{urlencode({'source': 'plan', 'plan_id': plan.id, 'q': item.folio})}",
        }

    def _build_document_focus_rows(
        items: list,
        row_builder,
        *,
        fallback_label: str = "",
        fallback_scope: str = "",
        fallback_count: int = 0,
        fallback_detail: str = "",
        fallback_action_label: str = "",
        fallback_action_url: str = "",
    ) -> list[dict]:
        rows = [row_builder(item) for item in items][:6]
        if rows or not fallback_count:
            return rows
        return [
            {
                "scope": fallback_scope,
                "folio": fallback_label,
                "status": f"{fallback_count} abiertos",
                "detail": fallback_detail,
                "action_label": fallback_action_label,
                "action_url": fallback_action_url,
            }
        ]

    stage_scope = "Solicitudes"
    stage_scope_url = f"{reverse('compras:solicitudes')}?{urlencode({'source': 'plan', 'plan_id': plan.id})}"
    stage_scope_detail = "Libera solicitudes y elimina bloqueos ERP para emitir órdenes."
    stage_scope_rows: list[dict] = []
    if current_stage in {"blocked_recepciones", "diferencias_abiertas", "recepciones_abiertas", "closed"}:
        stage_scope = "Recepciones"
        stage_scope_url = f"{reverse('compras:recepciones')}?{urlencode({'source': 'plan', 'plan_id': plan.id})}"
        stage_scope_rows = [_stage_focus_row_for_recepcion(item) for item in recepciones_items if getattr(item, "has_workflow_blockers", False)][:6]
        stage_scope_detail = (
            "Cierra recepciones con diferencias o bloqueos antes de terminar el ciclo."
            if stage_scope_rows or recepciones_abiertas
            else "Recepciones sin bloqueo operativo."
        )
    elif current_stage in {"blocked_ordenes", "ordenes_abiertas", "confirmed_ready"}:
        stage_scope = "Órdenes"
        stage_scope_url = f"{reverse('compras:ordenes')}?{urlencode({'source': 'plan', 'plan_id': plan.id})}"
        stage_scope_rows = [_stage_focus_row_for_orden(item) for item in ordenes_items if getattr(item, "has_workflow_blockers", False)][:6]
        stage_scope_detail = (
            "Confirma órdenes y elimina bloqueos documentales para abrir recepción."
            if stage_scope_rows or ordenes_abiertas
            else "Órdenes sin bloqueo operativo."
        )
    else:
        stage_scope_rows = [_stage_focus_row_for_solicitud(item) for item in solicitudes_items if getattr(item, "has_workflow_blockers", False)][:6]

    stage_focus = {
        "label": stage_scope,
        "detail": stage_scope_detail,
        "blocked_count": len(stage_scope_rows),
        "action_label": f"Abrir {stage_scope.lower()}",
        "action_url": stage_scope_url,
        "rows": stage_scope_rows,
        "tone": (
            "danger"
            if stage_scope_rows
            else "warning"
            if any(item["is_active"] and item["open"] > 0 for item in pipeline_steps)
            else "success"
            if any(item["is_active"] and item["closed"] > 0 for item in pipeline_steps)
            else "primary"
        ),
        "summary": (
            f"{len(stage_scope_rows)} bloqueos priorizados en {stage_scope.lower()}."
            if stage_scope_rows
            else f"{stage_scope} sin bloqueos priorizados."
        ),
    }
    closure_checks = [
        {
            "key": "solicitudes_liberadas",
            "label": "Solicitudes liberadas",
            "is_ready": solicitudes_pendientes == 0 and blocked_solicitudes == 0,
            "detail": "Sin pendientes de captura, validación ni bloqueos ERP.",
            "action_label": (
                "Revisar solicitudes bloqueadas"
                if blocked_solicitudes
                else "Liberar solicitudes pendientes"
                if solicitudes_pendientes
                else "Solicitudes al día"
            ),
            "action_url": f"{reverse('compras:solicitudes')}?{_plan_query({'closure_key': 'solicitudes_liberadas'})}",
            "action_detail": (
                "Corrige bloqueos ERP y completa datos faltantes antes de emitir órdenes."
                if blocked_solicitudes
                else "Termina captura o validación de solicitudes para liberar la etapa."
                if solicitudes_pendientes
                else "No quedan acciones abiertas en solicitudes para este plan."
            ),
        },
        {
            "key": "ordenes_sin_bloqueo",
            "label": "Órdenes sin bloqueo",
            "is_ready": blocked_ordenes == 0,
            "detail": "Proveedor, monto y datos documentales completos.",
            "action_label": (
                "Corregir órdenes bloqueadas"
                if blocked_ordenes
                else "Abrir órdenes del plan"
            ),
            "action_url": f"{reverse('compras:ordenes')}?{_plan_query({'closure_key': 'ordenes_sin_bloqueo'})}",
            "action_detail": (
                "Completa proveedor, monto estimado y datos documentales para habilitar recepción."
                if blocked_ordenes
                else "Las órdenes ya cumplen mínimos documentales para continuar el ciclo."
            ),
        },
        {
            "key": "recepciones_aplicadas",
            "label": "Recepciones aplicadas",
            "is_ready": recepciones_abiertas == 0 and blocked_recepciones == 0 and diferencias_abiertas == 0,
            "detail": "Sin recepciones abiertas, con diferencias o bloqueadas ERP.",
            "action_label": (
                "Resolver recepciones bloqueadas"
                if blocked_recepciones
                else "Atender diferencias abiertas"
                if diferencias_abiertas
                else "Cerrar recepciones abiertas"
                if recepciones_abiertas
                else "Recepciones al día"
            ),
            "action_url": f"{reverse('compras:recepciones')}?{_plan_query({'closure_key': 'recepciones_aplicadas'})}",
            "action_detail": (
                "Aplica correcciones ERP antes de cerrar recepción."
                if blocked_recepciones
                else "Justifica diferencias y normaliza conformidad para cerrar el plan."
                if diferencias_abiertas
                else "Confirma cantidades y aplica inventario para cerrar la etapa."
                if recepciones_abiertas
                else "No quedan recepciones abiertas ni diferencias pendientes."
            ),
        },
    ]
    for item in closure_checks:
        item["is_active"] = closure_key_filter == item["key"]

    handoff_checks = [
        {
            "key": "solicitud_orden",
            "label": "Solicitud → Orden",
            "is_ready": blocked_solicitudes == 0 and approved_ready == 0,
            "ready_count": approved_ready,
            "blocked_count": blocked_solicitudes,
            "detail": "Toda solicitud aprobada ya debe convertirse en orden o quedar bloqueada de forma explícita.",
            "action_label": (
                "Resolver solicitudes bloqueadas"
                if blocked_solicitudes
                else "Emitir órdenes pendientes"
                if approved_ready
                else "Handoff completo"
            ),
            "action_url": f"{reverse('compras:solicitudes')}?{_plan_query({'handoff_key': 'solicitud_orden'})}",
            "action_detail": (
                "Corrige maestro ERP o datos faltantes antes de emitir la orden."
                if blocked_solicitudes
                else "Genera órdenes para las solicitudes aprobadas todavía no convertidas."
                if approved_ready
                else "No quedan solicitudes listas sin convertir a orden."
            ),
        },
        {
            "key": "orden_recepcion",
            "label": "Orden → Recepción",
            "is_ready": blocked_ordenes == 0 and confirmed_ready == 0,
            "ready_count": confirmed_ready,
            "blocked_count": blocked_ordenes,
            "detail": "Toda orden confirmada debe pasar a recepción sin bloquear el siguiente tramo documental.",
            "action_label": (
                "Corregir órdenes bloqueadas"
                if blocked_ordenes
                else "Abrir recepciones pendientes"
                if confirmed_ready
                else "Handoff completo"
            ),
            "action_url": f"{reverse('compras:ordenes')}?{_plan_query({'handoff_key': 'orden_recepcion'})}",
            "action_detail": (
                "Completa proveedor, monto y trazabilidad para liberar la recepción."
                if blocked_ordenes
                else "Registra recepciones para órdenes confirmadas todavía abiertas."
                if confirmed_ready
                else "No quedan órdenes confirmadas pendientes de pasar a recepción."
            ),
        },
        {
            "key": "recepcion_cierre",
            "label": "Recepción → Cierre",
            "is_ready": blocked_recepciones == 0 and diferencias_abiertas == 0 and recepciones_abiertas == 0,
            "ready_count": recepciones_abiertas + diferencias_abiertas,
            "blocked_count": blocked_recepciones,
            "detail": "Toda recepción debe aplicarse a inventario y cerrar diferencias antes de concluir el plan.",
            "action_label": (
                "Resolver recepciones bloqueadas"
                if blocked_recepciones
                else "Cerrar recepciones abiertas"
                if recepciones_abiertas or diferencias_abiertas
                else "Handoff completo"
            ),
            "action_url": f"{reverse('compras:recepciones')}?{_plan_query({'handoff_key': 'recepcion_cierre'})}",
            "action_detail": (
                "Aplica correcciones ERP antes de cerrar recepción."
                if blocked_recepciones
                else "Aplica inventario y cierra diferencias para terminar el ciclo documental."
                if recepciones_abiertas or diferencias_abiertas
                else "La recepción del plan ya quedó cerrada documentalmente."
            ),
        },
    ]
    for item in handoff_checks:
        item["is_active"] = handoff_key_filter == item["key"]

    if current_stage in {"blocked_ordenes", "ordenes_abiertas", "confirmed_ready"}:
        preferred_handoff_index = 1
    elif current_stage in {"blocked_recepciones", "diferencias_abiertas", "recepciones_abiertas", "closed"}:
        preferred_handoff_index = 2
    else:
        preferred_handoff_index = 0
    active_handoff = next((item for item in handoff_checks if item["key"] == handoff_key_filter), None)
    handoff_focus = (
        active_handoff
        or (
            handoff_checks[preferred_handoff_index]
            if not handoff_checks[preferred_handoff_index]["is_ready"]
            else next((item for item in handoff_checks if not item["is_ready"]), None)
        )
    )
    if handoff_focus:
        handoff_focus = {
            **handoff_focus,
            "tone": "danger" if int(handoff_focus["blocked_count"]) > 0 else "warning",
            "summary": f"La entrega entre etapas sigue abierta por: {handoff_focus['label'].lower()}.",
        }
    else:
        handoff_focus = {
            "label": "Handoffs completos",
            "is_ready": True,
            "ready_count": 0,
            "blocked_count": 0,
            "detail": "Las entregas entre solicitudes, órdenes y recepciones ya quedaron completas.",
            "action_label": "Volver al plan",
            "action_url": f"{reverse('recetas:plan_produccion')}?{urlencode({'plan_id': plan.id})}",
            "action_detail": "No quedan handoffs documentales abiertos.",
            "tone": "success",
            "summary": "El flujo entre etapas ya está completo para este plan.",
        }
    if current_stage in {"blocked_ordenes", "ordenes_abiertas", "confirmed_ready"}:
        preferred_closure_index = 1
    elif current_stage in {"blocked_recepciones", "diferencias_abiertas", "recepciones_abiertas", "closed"}:
        preferred_closure_index = 2
    else:
        preferred_closure_index = 0
    active_closure = next((item for item in closure_checks if item["key"] == closure_key_filter), None)
    closure_focus = (
        active_closure
        or (
            closure_checks[preferred_closure_index]
            if not closure_checks[preferred_closure_index]["is_ready"]
            else next((item for item in closure_checks if not item["is_ready"]), None)
        )
    )
    if closure_focus:
        closure_focus = {
            **closure_focus,
            "tone": "danger" if current_stage in {"blocked_solicitudes", "blocked_ordenes", "blocked_recepciones"} else "warning",
            "summary": f"El cierre del plan sigue abierto por: {closure_focus['label'].lower()}.",
        }
    else:
        closure_focus = {
            "label": "Cierre documental completo",
            "is_ready": True,
            "detail": "Los tres criterios de cierre están completos para este plan.",
            "action_label": "Volver al plan",
            "action_url": f"{reverse('recetas:plan_produccion')}?{urlencode({'plan_id': plan.id})}",
            "action_detail": "No quedan bloqueos documentales abiertos en Compras.",
            "tone": "success",
            "summary": "El plan ya cerró criterios documentales en solicitudes, órdenes y recepciones.",
        }
    closure_focus_rows: list[dict] = []
    if closure_focus["label"] == "Solicitudes liberadas":
        closure_focus_rows = _build_document_focus_rows(
            [item for item in solicitudes_items if getattr(item, "has_workflow_blockers", False)],
            _stage_focus_row_for_solicitud,
            fallback_label="Solicitudes pendientes",
            fallback_scope="Solicitud",
            fallback_count=solicitudes_pendientes,
            fallback_detail="Solicitudes todavía en captura o validación dentro del plan.",
            fallback_action_label="Liberar solicitudes pendientes",
            fallback_action_url=f"{reverse('compras:solicitudes')}?{_plan_query({'estatus': SolicitudCompra.STATUS_BORRADOR})}",
        )
    elif closure_focus["label"] == "Órdenes sin bloqueo":
        closure_focus_rows = _build_document_focus_rows(
            [item for item in ordenes_items if getattr(item, "has_workflow_blockers", False)],
            _stage_focus_row_for_orden,
        )
    elif closure_focus["label"] == "Recepciones aplicadas":
        closure_focus_rows = _build_document_focus_rows(
            [item for item in recepciones_items if getattr(item, "has_workflow_blockers", False)],
            _stage_focus_row_for_recepcion,
            fallback_label="Recepciones abiertas",
            fallback_scope="Recepción",
            fallback_count=recepciones_abiertas + diferencias_abiertas,
            fallback_detail="Recepciones todavía abiertas o con diferencias por cerrar.",
            fallback_action_label="Cerrar recepciones abiertas",
            fallback_action_url=f"{reverse('compras:recepciones')}?{urlencode({'source': 'plan', 'plan_id': plan.id})}",
        )
    handoff_focus_rows: list[dict] = []
    if handoff_focus["label"] == "Solicitud → Orden":
        handoff_focus_rows = _build_document_focus_rows(
            [item for item in solicitudes_items if getattr(item, "has_workflow_blockers", False)],
            _stage_focus_row_for_solicitud,
            fallback_label="Solicitudes listas para OC",
            fallback_scope="Solicitud",
            fallback_count=approved_ready,
            fallback_detail="Solicitudes aprobadas todavía no convertidas en orden.",
            fallback_action_label="Emitir órdenes pendientes",
            fallback_action_url=f"{reverse('compras:solicitudes')}?{_plan_query({'estatus': 'APPROVED_READY'})}",
        )
    elif handoff_focus["label"] == "Orden → Recepción":
        handoff_focus_rows = _build_document_focus_rows(
            [item for item in ordenes_items if getattr(item, "has_workflow_blockers", False)],
            _stage_focus_row_for_orden,
            fallback_label="Órdenes listas para recepción",
            fallback_scope="Orden",
            fallback_count=confirmed_ready,
            fallback_detail="Órdenes confirmadas todavía sin recepción abierta.",
            fallback_action_label="Abrir recepciones pendientes",
            fallback_action_url=f"{reverse('compras:ordenes')}?{_plan_query({'estatus': OrdenCompra.STATUS_CONFIRMADA})}",
        )
    elif handoff_focus["label"] == "Recepción → Cierre":
        handoff_focus_rows = _build_document_focus_rows(
            [item for item in recepciones_items if getattr(item, "has_workflow_blockers", False)],
            _stage_focus_row_for_recepcion,
            fallback_label="Recepciones por cerrar",
            fallback_scope="Recepción",
            fallback_count=recepciones_abiertas + diferencias_abiertas,
            fallback_detail="Recepciones todavía sin aplicar o con diferencias abiertas.",
            fallback_action_label="Cerrar recepciones abiertas",
            fallback_action_url=f"{reverse('compras:recepciones')}?{_plan_query()}",
        )

    solicitudes_master_rows = [
        {
            "folio": solicitud.folio,
            "details": _enterprise_blocker_details_for_solicitud(solicitud),
        }
        for solicitud in solicitudes_items
        if _enterprise_blocker_details_for_solicitud(solicitud)
    ]
    ordenes_master_rows = [
        {
            "folio": orden.folio,
            "details": _enterprise_blocker_details_for_solicitud(orden.solicitud),
        }
        for orden in ordenes_items
        if orden.solicitud_id and _enterprise_blocker_details_for_solicitud(orden.solicitud)
    ]
    recepciones_master_rows = [
        {
            "folio": recepcion.folio,
            "details": _enterprise_blocker_details_for_solicitud(recepcion.orden.solicitud),
        }
        for recepcion in recepciones_items
        if recepcion.orden_id and getattr(recepcion.orden, "solicitud_id", None) and _enterprise_blocker_details_for_solicitud(recepcion.orden.solicitud)
    ]
    if current_stage in {"blocked_recepciones", "diferencias_abiertas", "recepciones_abiertas", "closed"}:
        preferred_master_rows = recepciones_master_rows
    elif current_stage in {"blocked_ordenes", "ordenes_abiertas", "confirmed_ready"}:
        preferred_master_rows = ordenes_master_rows
    else:
        preferred_master_rows = solicitudes_master_rows
    master_rollup = _enterprise_master_blocker_rollup(
        preferred_master_rows or solicitudes_master_rows or ordenes_master_rows or recepciones_master_rows,
        usage_scope="purchases",
    )
    master_blocker_class_cards = list(master_rollup.get("master_blocker_class_cards") or [])
    master_blocker_missing_cards = list(master_rollup.get("master_blocker_missing_cards") or [])
    master_blocker_detail_rows = list(master_rollup.get("master_blocker_detail_rows") or [])
    for card in master_blocker_class_cards:
        card["is_active"] = master_class_filter == card["class_key"]
        card["focus_url"] = f"{reverse(f'compras:{current_view}')}?{_plan_query({'master_class': card['class_key']})}"
    for card in master_blocker_missing_cards:
        card["is_active"] = master_missing_filter == card["key"]
        card["focus_url"] = f"{reverse(f'compras:{current_view}')}?{_plan_query({'master_missing': card['key']})}"
    if master_class_filter != "all":
        master_blocker_detail_rows = [
            row for row in master_blocker_detail_rows if row.get("class_key") == master_class_filter
        ]
    if master_missing_filter != "all":
        master_blocker_detail_rows = [
            row for row in master_blocker_detail_rows if _missing_field_filter_key(row.get("missing_field")) == master_missing_filter
        ]
    master_focus_rows = master_blocker_detail_rows[:3]
    if master_focus_rows:
        first_master_focus = master_focus_rows[0]
        master_focus = {
            **first_master_focus,
            "label": f"{first_master_focus['class_label']} · {first_master_focus['missing_field']}",
            "summary": (
                f"El flujo del plan sigue bloqueado por {first_master_focus['insumo_nombre']} "
                f"({first_master_focus['missing_field']})."
            ),
            "tone": "danger" if first_master_focus.get("tone") == "danger" else "warning",
        }
        if master_class_filter != "all" or master_missing_filter != "all":
            master_focus["summary"] = (
                f"Vista enfocada en {first_master_focus['class_label'].lower()} · "
                f"{first_master_focus['missing_field'].lower()}."
            )
    else:
        master_focus = {
            "label": "Maestro ERP al día",
            "summary": "No se detectan bloqueos prioritarios del maestro para este plan.",
            "action_label": "Abrir maestro",
            "action_url": reverse("maestros:insumo_list"),
            "action_detail": "El catálogo no muestra bloqueos maestros sobre los documentos del plan.",
            "tone": "success",
            "class_label": "",
            "missing_field": "",
            "insumo_nombre": "",
            "count": 0,
        }

    document_ready_checks = sum(1 for item in closure_checks if item.get("is_ready"))
    document_total_checks = len(closure_checks)
    document_progress_pct = int(round((document_ready_checks / document_total_checks) * 100)) if document_total_checks else 100
    handoff_ready_checks = sum(1 for item in handoff_checks if item.get("is_ready"))
    handoff_total_checks = len(handoff_checks)
    handoff_progress_pct = int(round((handoff_ready_checks / handoff_total_checks) * 100)) if handoff_total_checks else 100

    def _completion_for_gate(*, ready: bool, blocked: int = 0, pending: int = 0) -> int:
        if ready:
            return 100
        if blocked > 0:
            return 15
        if pending > 0:
            return 50
        return 0

    if current_view == "recepciones":
        dependency_gate = {
            "label": "Órdenes sin bloqueo",
            "owner": "Compras / Recepción",
            "is_ready": blocked_ordenes == 0,
            "blocked": blocked_ordenes,
            "pending": confirmed_ready,
            "detail": "Recepciones dependen de órdenes confirmadas, completas y trazables.",
            "next_step": (
                "Corrige proveedor, monto y trazabilidad de la orden antes de recibir."
                if blocked_ordenes
                else "Convierte órdenes confirmadas en recepciones aplicadas."
                if confirmed_ready
                else "Las órdenes previas ya están listas para recepción."
            ),
            "url": f"{reverse('compras:ordenes')}?{_plan_query({'handoff_key': 'orden_recepcion'})}",
            "cta": (
                "Corregir órdenes bloqueadas"
                if blocked_ordenes
                else "Abrir órdenes confirmadas"
                if confirmed_ready
                else "Abrir órdenes del plan"
            ),
            "depends_on": "Solicitudes liberadas y órdenes confirmadas",
            "exit_criteria": "Órdenes confirmadas sin bloqueo ERP y listas para recibirse.",
        }
    elif current_view == "ordenes":
        dependency_gate = {
            "label": "Solicitudes liberadas",
            "owner": "Compras / Abastecimiento",
            "is_ready": blocked_solicitudes == 0 and solicitudes_pendientes == 0,
            "blocked": blocked_solicitudes,
            "pending": solicitudes_pendientes,
            "detail": "Órdenes dependen de solicitudes aprobadas, trazables y sin bloqueo ERP.",
            "next_step": (
                "Corrige artículo, costo o proveedor antes de emitir la orden."
                if blocked_solicitudes
                else "Libera solicitudes pendientes para habilitar órdenes."
                if solicitudes_pendientes
                else "Las solicitudes previas ya están listas para emisión documental."
            ),
            "url": f"{reverse('compras:solicitudes')}?{_plan_query({'handoff_key': 'solicitud_orden'})}",
            "cta": (
                "Resolver solicitudes bloqueadas"
                if blocked_solicitudes
                else "Liberar solicitudes"
                if solicitudes_pendientes
                else "Abrir solicitudes del plan"
            ),
            "depends_on": "Plan cargado y solicitudes aprobadas",
            "exit_criteria": "Solicitudes del plan aprobadas y sin bloqueo ERP.",
        }
    else:
        dependency_gate = {
            "label": "Plan y demanda operativa",
            "owner": "Producción / Planeación",
            "is_ready": blocked_total == 0,
            "blocked": blocked_total,
            "pending": solicitudes_pendientes,
            "detail": "Solicitudes parten del plan, la demanda operativa y el maestro vigente.",
            "next_step": stage_detail,
            "url": f"{reverse('recetas:plan_produccion')}?{urlencode({'plan_id': plan.id})}",
            "cta": "Abrir plan",
            "depends_on": "Recetas, BOM, demanda y fechas vigentes",
            "exit_criteria": "Plan documentado y necesidades listas para captura en compras.",
        }

    demand_signal = _build_plan_demand_signal(plan, session=session)
    demand_gate = _build_plan_demand_gate(demand_signal)
    branch_priority_rows = _build_plan_branch_priority_rows(
        plan,
        periodo=plan.fecha_produccion.strftime("%Y-%m"),
    )
    branch_supply_rows = _build_plan_branch_supply_rows(plan, branch_priority_rows)
    commercial_priority_rows = _build_plan_commercial_priority_rows(plan)
    master_demand_gate = _build_plan_master_demand_gate(commercial_priority_rows)
    demand_dependency_row = None
    if demand_signal:
        if demand_signal["forecast_count"]:
            demand_status = demand_signal["alignment_label"]
            demand_tone = (
                "success"
                if demand_signal["alignment_pct"] >= 80
                else "warning"
                if demand_signal["alignment_pct"] >= 50
                else "danger"
            )
            demand_completion = demand_signal["alignment_pct"]
            demand_detail = (
                f"Forecast activo en {demand_signal['forecast_count']} recetas del plan con "
                f"{demand_signal['avg_confidence']:.1f}% de confianza promedio."
            )
            demand_next_step = "Ajusta solicitud y abastecimiento contra el forecast comparable antes de emitir compras."
        else:
            demand_status = demand_signal["historico_status"]
            demand_tone = demand_signal["historico_tone"]
            demand_completion = min(demand_signal["historico_days"] * 3, 100)
            demand_detail = demand_signal["historico_detail"]
            demand_next_step = "Construye forecast o valida cobertura histórica antes de escalar solicitudes del plan."
        demand_dependency_row = {
            "label": "Demanda histórica / forecast",
            "owner": "Planeación comercial",
            "status": demand_status,
            "tone": demand_tone,
            "blockers": int(demand_gate["blockers"]),
            "completion": demand_completion,
            "depends_on": "Ventas históricas, forecast y solicitud de ventas",
            "exit_criteria": "Base histórica comparable y solicitud alineada para abastecimiento.",
            "detail": demand_detail,
            "next_step": demand_next_step,
            "url": f"{reverse('recetas:plan_produccion')}?{urlencode({'plan_id': plan.id})}",
            "cta": "Abrir plan",
        }

    upstream_dependency_rows = [
        {
            "label": "Plan de producción",
            "owner": "Producción / Planeación",
            "status": "Listo" if blocked_total == 0 else stage_label,
            "tone": "success" if blocked_total == 0 else stage_tone,
            "blockers": blocked_total,
            "completion": max(document_progress_pct, handoff_progress_pct) if blocked_total else 100,
            "depends_on": "Recetas, BOM y calendario productivo",
            "exit_criteria": "Plan con documentos trazables y flujo operativo definido.",
            "detail": stage_detail,
            "next_step": next_action["detail"],
            "url": f"{reverse('recetas:plan_produccion')}?{urlencode({'plan_id': plan.id})}",
            "cta": "Abrir plan",
        },
        {
            "label": "Maestro de artículos",
            "owner": "Maestros / Compras",
            "status": (
                master_demand_gate["status"]
                if int(master_demand_gate["blockers"]) > 0
                else "Listo" if not master_focus_rows else master_focus["label"]
            ),
            "tone": (
                master_demand_gate["tone"]
                if int(master_demand_gate["blockers"]) > 0
                else "success" if not master_focus_rows else master_focus["tone"]
            ),
            "blockers": (
                int(master_demand_gate["blockers"])
                if int(master_demand_gate["blockers"]) > 0
                else len(master_focus_rows)
            ),
            "completion": (
                25
                if str(master_demand_gate["tone"]) == "danger"
                else 60
                if int(master_demand_gate["blockers"]) > 0
                else 100 if not master_focus_rows else 0
            ),
            "depends_on": "Unidad base, proveedor, categoría y código comercial",
            "exit_criteria": "Artículos requeridos sin bloqueos ERP.",
            "detail": (
                master_demand_gate["detail"]
                if int(master_demand_gate["blockers"]) > 0
                else master_focus["summary"]
            ),
            "next_step": (
                master_demand_gate["next_step"]
                if int(master_demand_gate["blockers"]) > 0
                else master_focus["action_detail"]
            ),
            "url": (
                master_demand_gate["action_url"]
                if int(master_demand_gate["blockers"]) > 0
                else master_focus["action_url"]
            ),
            "cta": (
                master_demand_gate["action_label"]
                if int(master_demand_gate["blockers"]) > 0
                else master_focus["action_label"]
            ),
        },
        {
            "label": dependency_gate["label"],
            "owner": dependency_gate["owner"],
            "status": "Listo" if dependency_gate["is_ready"] else "En revisión",
            "tone": "success" if dependency_gate["is_ready"] else ("danger" if dependency_gate["blocked"] > 0 else "warning"),
            "blockers": dependency_gate["blocked"],
            "completion": _completion_for_gate(
                ready=bool(dependency_gate["is_ready"]),
                blocked=int(dependency_gate["blocked"]),
                pending=int(dependency_gate["pending"]),
            ),
            "depends_on": dependency_gate["depends_on"],
            "exit_criteria": dependency_gate["exit_criteria"],
            "detail": dependency_gate["detail"],
            "next_step": dependency_gate["next_step"],
            "url": dependency_gate["url"],
            "cta": dependency_gate["cta"],
        },
    ]
    if demand_dependency_row:
        upstream_dependency_rows.insert(1, demand_dependency_row)

    master_demand_blocks_issue = (
        str(master_demand_gate.get("tone") or "") == "danger"
        and int(master_demand_gate.get("blockers") or 0) > 0
    )
    if master_demand_blocks_issue:
        stage_label = "Demanda crítica bloqueada"
        stage_tone = "danger"
        stage_detail = str(master_demand_gate.get("detail") or stage_detail)
        next_action = {
            "label": str(master_demand_gate.get("action_label") or "Cerrar artículos críticos"),
            "detail": str(master_demand_gate.get("next_step") or "Completa el maestro de los artículos críticos antes de continuar."),
            "url": str(master_demand_gate.get("action_url") or reverse("maestros:insumo_list")),
        }
    daily_critical_close_focus = None
    critical_master_demand_rows = list(master_demand_gate.get("rows") or [])[:3]
    if critical_master_demand_rows:
        top_row = critical_master_demand_rows[0]
        daily_critical_close_focus = {
            "title": "Cierre prioritario del día",
            "detail": (
                f"{top_row['insumo_nombre']} debe cerrarse primero para liberar el plan y la emisión documental."
            ),
            "historico_units": top_row["historico_units"],
            "required_qty": top_row["required_qty"],
            "missing": top_row["master_missing"],
            "url": top_row["action_url"],
            "cta": "Cerrar artículo ahora",
            "tone": "danger",
        }

    return {
        "plan": plan,
        "plan_id": plan.id,
        "plan_scope": plan_scope,
        "label": plan.nombre or f"Plan {plan.id}",
        "fecha_produccion": plan.fecha_produccion,
        "back_url": f"{reverse('recetas:plan_produccion')}?{urlencode({'plan_id': plan.id})}",
        "solicitudes_url": f"{reverse('compras:solicitudes')}?{solicitudes_query}",
        "ordenes_url": f"{reverse('compras:ordenes')}?{solicitudes_query}",
        "recepciones_url": f"{reverse('compras:recepciones')}?{solicitudes_query}",
        "blocked_total": blocked_total,
        "summary_label": summary_label,
        "summary_cards": summary_cards,
        "stage_label": stage_label,
        "stage_tone": stage_tone,
        "stage_detail": stage_detail,
        "next_action": next_action,
        "drilldown_actions": drilldown_actions,
        "pipeline_steps": pipeline_steps,
        "stage_focus": stage_focus,
        "closure_checks": closure_checks,
        "document_ready_checks": document_ready_checks,
        "document_total_checks": document_total_checks,
        "document_progress_pct": document_progress_pct,
        "closure_focus": closure_focus,
        "closure_focus_rows": closure_focus_rows,
        "handoff_checks": handoff_checks,
        "handoff_ready_checks": handoff_ready_checks,
        "handoff_total_checks": handoff_total_checks,
        "handoff_progress_pct": handoff_progress_pct,
        "handoff_focus": handoff_focus,
        "handoff_focus_rows": handoff_focus_rows,
        "master_blocker_class_cards": master_blocker_class_cards,
        "master_blocker_missing_cards": master_blocker_missing_cards,
        "master_focus": master_focus,
        "master_focus_rows": master_focus_rows,
        "upstream_dependency_rows": upstream_dependency_rows,
        "demand_signal": demand_signal,
        "demand_gate": demand_gate,
        "branch_priority_rows": branch_priority_rows,
        "branch_supply_rows": branch_supply_rows,
        "commercial_priority_rows": commercial_priority_rows,
        "master_demand_gate": master_demand_gate,
        "critical_master_demand_rows": critical_master_demand_rows,
        "master_demand_blocks_issue": master_demand_blocks_issue,
        "daily_critical_close_focus": daily_critical_close_focus,
    }


def _build_plan_demand_signal(plan: PlanProduccion, *, session: dict | None = None) -> dict | None:
    receta_ids = list(plan.items.values_list("receta_id", flat=True).distinct())
    if not receta_ids:
        return None

    historico_qs = VentaHistorica.objects.filter(
        receta_id__in=receta_ids,
        fecha__lt=plan.fecha_produccion,
    )
    historico_total = historico_qs.aggregate(total=Sum("cantidad")).get("total") or Decimal("0")
    historico_rows = historico_qs.count()
    historico_days = historico_qs.values("fecha").distinct().count()
    historico_years = historico_qs.dates("fecha", "year").count()
    comparable_years = (
        historico_qs.filter(fecha__month=plan.fecha_produccion.month).dates("fecha", "year").count()
    )
    historico_sucursales = historico_qs.exclude(sucursal_id=None).values("sucursal_id").distinct().count()
    historico_recetas = historico_qs.values("receta_id").distinct().count()
    top_historial = list(
        historico_qs.values("receta__nombre").annotate(total=Sum("cantidad")).order_by("-total", "receta__nombre")[:5]
    )
    if historico_days >= 28 and comparable_years >= 3:
        historico_status = "Base robusta multianual"
        historico_tone = "success"
        historico_detail = "El plan ya tiene base histórica suficiente y comparable en varios años para orientar abastecimiento."
    elif historico_days >= 28:
        historico_status = "Base robusta"
        historico_tone = "success"
        historico_detail = "El plan ya tiene base histórica suficiente para orientar abastecimiento."
    elif historico_days >= 7 and historico_years >= 2:
        historico_status = "Base utilizable multianual"
        historico_tone = "warning"
        historico_detail = "La demanda histórica del plan ya cubre más de un año, aunque todavía requiere criterio operativo."
    elif historico_days >= 7:
        historico_status = "Base utilizable"
        historico_tone = "warning"
        historico_detail = "La demanda histórica ya orienta el plan, aunque todavía requiere criterio operativo."
    else:
        historico_status = "Base limitada"
        historico_tone = "danger"
        historico_detail = "La cobertura histórica del plan sigue siendo corta para empujar compras con confianza."

    preview_payload = (session or {}).get("pronostico_estadistico_preview") or {}
    preview_rows = []
    for row in list(preview_payload.get("rows") or []):
        try:
            receta_id = int(row.get("receta_id") or 0)
        except Exception:
            receta_id = 0
        if receta_id in receta_ids:
            preview_rows.append(row)
    forecast_count = len(preview_rows)
    avg_confidence = Decimal("0")
    top_forecast = []
    if preview_rows:
        avg_confidence = (
            sum((Decimal(str(row.get("confianza") or 0)) for row in preview_rows), Decimal("0"))
            / Decimal(str(forecast_count))
        ).quantize(Decimal("0.1"))
        top_forecast = sorted(
            preview_rows,
            key=lambda row: Decimal(str(row.get("forecast_qty") or 0)),
            reverse=True,
        )[:5]

    solicitud_map = {
        int(row["receta_id"]): Decimal(str(row["total"] or 0))
        for row in (
            SolicitudVenta.objects.filter(
                receta_id__in=receta_ids,
                fecha_inicio__lte=plan.fecha_produccion,
                fecha_fin__gte=plan.fecha_produccion,
            )
            .values("receta_id")
            .annotate(total=Sum("cantidad"))
        )
    }
    aligned = 0
    evaluated = 0
    deviation_rows: list[dict[str, Any]] = []
    for row in preview_rows:
        receta_id = int(row.get("receta_id") or 0)
        forecast_qty = Decimal(str(row.get("forecast_qty") or 0))
        solicitud_qty = solicitud_map.get(receta_id, Decimal("0"))
        if forecast_qty <= 0 and solicitud_qty <= 0:
            continue
        evaluated += 1
        tolerance = max(Decimal("1"), forecast_qty * Decimal("0.10"))
        delta = solicitud_qty - forecast_qty
        if abs(delta) <= tolerance:
            aligned += 1
        deviation_rows.append(
            {
                "receta": row.get("receta") or "",
                "forecast_qty": forecast_qty,
                "solicitud_qty": solicitud_qty,
                "delta_qty": delta,
            }
        )
    alignment_pct = int(round((aligned / evaluated) * 100)) if evaluated else 0
    top_deviations = sorted(deviation_rows, key=lambda row: abs(row["delta_qty"]), reverse=True)[:5]
    if evaluated == 0:
        alignment_label = "Sin solicitud comparable"
    elif alignment_pct >= 80:
        alignment_label = "Alineación sólida"
    elif alignment_pct >= 50:
        alignment_label = "Alineación parcial"
    else:
        alignment_label = "Alineación frágil"

    return {
        "available": historico_rows > 0 or forecast_count > 0,
        "historico_status": historico_status,
        "historico_tone": historico_tone,
        "historico_detail": historico_detail,
        "historico_rows": historico_rows,
        "historico_days": historico_days,
        "historico_years": historico_years,
        "comparable_years": comparable_years,
        "historico_sucursales": historico_sucursales,
        "historico_recetas": historico_recetas,
        "historico_total": historico_total,
        "top_historial": top_historial,
        "forecast_count": forecast_count,
        "avg_confidence": avg_confidence,
        "top_forecast": top_forecast,
        "alignment_pct": alignment_pct,
        "alignment_label": alignment_label,
        "alignment_evaluated": evaluated,
        "top_deviations": top_deviations,
    }


def _build_plan_demand_gate(signal: dict | None) -> dict[str, object]:
    if not signal or not bool(signal.get("available")):
        return {
            "status": "Sin base comercial",
            "tone": "danger",
            "is_ready": False,
            "blockers": 1,
            "detail": "Todavía no hay demanda histórica ni forecast suficiente para respaldar compras del plan.",
            "next_step": "Revisar forecast y demanda histórica antes de liberar documentos.",
        }
    tone = str(signal.get("historico_tone") or "warning")
    years_observed = int(signal.get("historico_years") or 0)
    if tone == "success":
        return {
            "status": str(signal.get("historico_status") or "Base comercial lista"),
            "tone": "success",
            "is_ready": True,
            "blockers": 0,
            "detail": str(signal.get("historico_detail") or ""),
            "next_step": "La base comercial ya puede respaldar solicitudes, órdenes y recepciones.",
        }
    if tone == "warning" and years_observed >= 2:
        return {
            "status": str(signal.get("historico_status") or "Base comercial en revisión"),
            "tone": "warning",
            "is_ready": True,
            "blockers": 0,
            "detail": str(signal.get("historico_detail") or ""),
            "next_step": "Opera compras con criterio y valida el comparativo forecast/solicitud.",
        }
    return {
        "status": "Base comercial frágil",
        "tone": "danger",
        "is_ready": False,
        "blockers": 1,
        "detail": str(signal.get("historico_detail") or "La base comercial todavía es débil para escalar compras con confianza."),
        "next_step": "Refuerza histórico, forecast o solicitud antes de cerrar el flujo documental.",
    }


def _build_plan_branch_priority_rows(plan: PlanProduccion, *, periodo: str, limit: int = 6) -> list[dict[str, object]]:
    plan_items = list(plan.items.select_related("receta").all())
    if not plan_items:
        return []

    receta_ids = [item.receta_id for item in plan_items if item.receta_id]
    if not receta_ids:
        return []

    month = None
    try:
        month = int((periodo or "").split("-")[1])
    except (IndexError, TypeError, ValueError):
        month = None

    sales_qs = VentaHistorica.objects.filter(receta_id__in=receta_ids, sucursal_id__isnull=False)
    if month:
        monthly_qs = sales_qs.filter(fecha__month=month)
        sales_qs = monthly_qs if monthly_qs.exists() else sales_qs

    sales_rows = list(
        sales_qs.values("sucursal_id", "sucursal__codigo", "sucursal__nombre")
        .annotate(total_units=Sum("cantidad"), recipe_count=Count("receta_id", distinct=True))
        .order_by("-total_units", "sucursal__codigo")
    )
    if not sales_rows:
        return []

    request_map = {
        int(row["sucursal_id"]): Decimal(str(row["solicitud_total"] or 0))
        for row in (
            SolicitudVenta.objects.filter(
                periodo=periodo,
                receta_id__in=receta_ids,
                sucursal_id__isnull=False,
            )
            .values("sucursal_id")
            .annotate(solicitud_total=Sum("cantidad"))
        )
        if row.get("sucursal_id")
    }
    recipe_plan_map = {int(item.receta_id): Decimal(str(item.cantidad or 0)) for item in plan_items}
    branch_recipe_rows = list(
        sales_qs.values("sucursal_id", "receta_id", "receta__nombre")
        .annotate(total_units=Sum("cantidad"))
        .order_by("sucursal_id", "-total_units", "receta__nombre")
    )
    dominant_recipe_map: dict[int, dict[str, object]] = {}
    for row in branch_recipe_rows:
        sucursal_id = int(row.get("sucursal_id") or 0)
        receta_id = int(row.get("receta_id") or 0)
        if not sucursal_id or not receta_id or sucursal_id in dominant_recipe_map:
            continue
        dominant_recipe_map[sucursal_id] = {
            "recipe_id": receta_id,
            "recipe_name": row.get("receta__nombre") or "Producto",
            "recipe_units": Decimal(str(row.get("total_units") or 0)),
            "plan_qty": recipe_plan_map.get(receta_id, Decimal("0")),
        }

    rows: list[dict[str, object]] = []
    for row in sales_rows:
        sucursal_id = int(row.get("sucursal_id") or 0)
        if not sucursal_id:
            continue
        historico_units = Decimal(str(row.get("total_units") or 0))
        solicitud_total = request_map.get(sucursal_id, Decimal("0"))
        recipe_count = int(row.get("recipe_count") or 0)
        dominant_recipe = dominant_recipe_map.get(sucursal_id) or {}
        dominant_recipe_id = int(dominant_recipe.get("recipe_id") or 0)
        dominant_recipe_name = str(dominant_recipe.get("recipe_name") or "Producto")
        dominant_recipe_units = Decimal(str(dominant_recipe.get("recipe_units") or 0))
        dominant_plan_qty = Decimal(str(dominant_recipe.get("plan_qty") or 0))

        if solicitud_total > 0 and historico_units <= 0:
            tone = "danger"
            status = "Solicitud sin base"
            detail = "La sucursal ya empuja compra, pero no trae respaldo comparable suficiente en el periodo."
            priority_score = solicitud_total * Decimal("10")
            action_label = "Revisar demanda"
        elif solicitud_total > historico_units and historico_units > 0:
            tone = "warning"
            status = "Presión superior al histórico"
            detail = f"Solicitud {solicitud_total:.0f} vs histórico comparable {historico_units:.0f}."
            priority_score = (solicitud_total * Decimal("8")) + historico_units
            action_label = "Alinear presión"
        elif solicitud_total > 0:
            tone = "primary"
            status = "Solicitud activa"
            detail = f"La sucursal ya empuja {solicitud_total:.0f} unidades para {recipe_count} producto(s) del plan."
            priority_score = (solicitud_total * Decimal("6")) + historico_units
            action_label = "Abrir plan"
        else:
            tone = "success"
            status = "Demanda comparable"
            detail = f"Trae {historico_units:.0f} unidades comparables para {recipe_count} producto(s) del plan."
            priority_score = historico_units
            action_label = "Ver plan"

        rows.append(
            {
                "sucursal_codigo": row.get("sucursal__codigo") or "",
                "sucursal_nombre": row.get("sucursal__nombre") or "Sucursal",
                "status": status,
                "tone": tone,
                "detail": detail,
                "historico_units": historico_units,
                "solicitud_total": solicitud_total,
                "recipe_count": recipe_count,
                "dominant_recipe_id": dominant_recipe_id,
                "dominant_recipe_name": dominant_recipe_name,
                "dominant_recipe_units": dominant_recipe_units,
                "dominant_plan_qty": dominant_plan_qty,
                "action_url": f"{reverse('recetas:plan_produccion')}?{urlencode({'plan_id': plan.id, 'periodo': periodo})}#plan-productos",
                "action_label": action_label,
                "priority_score": priority_score,
            }
        )

    tone_order = {"danger": 0, "warning": 1, "primary": 2, "success": 3}
    rows.sort(
        key=lambda item: (
            tone_order.get(str(item.get("tone") or ""), 9),
            -float(item.get("priority_score") or 0),
            str(item.get("sucursal_codigo") or ""),
        )
    )
    return rows[:limit]


def _build_plan_branch_supply_rows(
    plan: PlanProduccion,
    branch_priority_rows: list[dict[str, object]],
    *,
    limit: int = 6,
) -> list[dict[str, object]]:
    if not branch_priority_rows:
        return []

    recipe_ids = {
        int(row.get("dominant_recipe_id") or 0)
        for row in branch_priority_rows
        if int(row.get("dominant_recipe_id") or 0) > 0
    }
    if not recipe_ids:
        return []

    lineas = list(
        LineaReceta.objects.filter(receta_id__in=recipe_ids, insumo_id__isnull=False)
        .exclude(tipo_linea=LineaReceta.TIPO_SUBSECCION)
        .select_related("insumo", "receta")
    )
    if not lineas:
        return []

    canonical_ids: set[int] = set()
    canonical_by_line: dict[int, Insumo] = {}
    for linea in lineas:
        if not linea.insumo:
            continue
        canonical = canonical_insumo(linea.insumo) or linea.insumo
        canonical_by_line[linea.id] = canonical
        canonical_ids.add(canonical.id)

    existencia_map = {
        int(item.insumo_id): item
        for item in ExistenciaInsumo.objects.filter(insumo_id__in=canonical_ids).select_related("insumo")
    }
    lineas_by_recipe: dict[int, list[LineaReceta]] = defaultdict(list)
    for linea in lineas:
        lineas_by_recipe[int(linea.receta_id)].append(linea)

    rows: list[dict[str, object]] = []
    for branch_row in branch_priority_rows:
        recipe_id = int(branch_row.get("dominant_recipe_id") or 0)
        plan_qty = Decimal(str(branch_row.get("dominant_plan_qty") or 0))
        if not recipe_id or plan_qty <= 0:
            continue

        best_candidate: dict[str, object] | None = None
        best_score = Decimal("-1")
        for linea in lineas_by_recipe.get(recipe_id, []):
            canonical = canonical_by_line.get(linea.id)
            if canonical is None:
                continue
            required_qty = Decimal(str(linea.cantidad or 0)) * plan_qty
            if required_qty <= 0:
                continue
            existencia = existencia_map.get(canonical.id)
            stock_actual = Decimal(str(getattr(existencia, "stock_actual", 0) or 0))
            shortage = max(required_qty - stock_actual, Decimal("0"))
            readiness = enterprise_readiness_profile(canonical)
            missing = list(readiness.get("missing") or [])
            latest_cost = latest_costo_canonico(insumo_id=canonical.id)
            missing_cost = latest_cost is None
            score = (shortage * Decimal("100")) + (Decimal(str(len(missing))) * Decimal("50")) + required_qty
            if missing_cost:
                score += Decimal("25")
            if score > best_score:
                best_score = score
                best_candidate = {
                    "insumo_id": canonical.id,
                    "insumo_nombre": canonical.nombre,
                    "required_qty": required_qty,
                    "stock_actual": stock_actual,
                    "shortage": shortage,
                    "master_missing": missing,
                    "missing_cost": missing_cost,
                    "unidad": canonical.unidad_base.codigo if getattr(canonical, "unidad_base_id", None) else "-",
                    "action_url": reverse("maestros:insumo_update", args=[canonical.id]),
                }

        if not best_candidate:
            continue

        rows.append(
            {
                "sucursal_codigo": branch_row.get("sucursal_codigo") or "",
                "sucursal_nombre": branch_row.get("sucursal_nombre") or "Sucursal",
                "dominant_recipe_name": branch_row.get("dominant_recipe_name") or "Producto",
                "insumo_nombre": best_candidate["insumo_nombre"],
                "required_qty": best_candidate["required_qty"],
                "stock_actual": best_candidate["stock_actual"],
                "shortage": best_candidate["shortage"],
                "master_missing": best_candidate["master_missing"],
                "missing_cost": best_candidate["missing_cost"],
                "unidad": best_candidate["unidad"],
                "action_url": best_candidate["action_url"],
                "action_label": "Asegurar insumo",
                "priority_score": best_score,
            }
        )

    rows.sort(
        key=lambda item: (
            -float(item.get("shortage") or 0),
            -len(item.get("master_missing") or []),
            -float(item.get("required_qty") or 0),
            str(item.get("sucursal_codigo") or ""),
        )
    )
    return rows[:limit]


def _build_plan_master_demand_gate(rows: list[dict[str, object]] | None) -> dict[str, object]:
    rows = rows or []
    critical_rows = [row for row in rows if str(row.get("priority_tone") or "") == "danger"]
    if critical_rows:
        top_row = critical_rows[0]
        return {
            "status": "Demanda crítica bloqueada por maestro",
            "tone": "danger",
            "is_ready": False,
            "blockers": len(critical_rows),
            "rows": critical_rows[:3],
            "detail": (
                f"{len(critical_rows)} artículo(s) del plan ya sostienen demanda fuerte y siguen incompletos en maestro. "
                f"Prioridad actual: {top_row.get('insumo_nombre', 'Artículo')}."
            ),
            "next_step": "Cierra primero el artículo maestro crítico antes de emitir o recibir documentos del plan.",
            "action_label": "Cerrar prioridad crítica",
            "action_url": str(top_row.get("action_url") or reverse("maestros:insumo_list")),
        }
    if rows:
        top_row = rows[0]
        return {
            "status": "Demanda priorizada en revisión",
            "tone": "warning",
            "is_ready": False,
            "blockers": len(rows),
            "rows": rows[:3],
            "detail": "Hay artículos del plan con demanda relevante y faltantes en maestro que conviene cerrar antes de escalar abastecimiento.",
            "next_step": "Revisa primero los artículos del plan con mayor demanda antes de cerrar órdenes o recepciones.",
            "action_label": "Abrir artículos del plan",
            "action_url": str(top_row.get("action_url") or reverse("maestros:insumo_list")),
        }
    return {
        "status": "Maestro comercial del plan estable",
        "tone": "success",
        "is_ready": True,
        "blockers": 0,
        "rows": [],
        "detail": "No hay artículos de alta demanda del plan bloqueados por faltantes del maestro.",
        "next_step": "Mantén seguimiento preventivo del maestro mientras avanzan los documentos.",
        "action_label": "Abrir maestro",
        "action_url": reverse("maestros:insumo_list"),
    }


def _build_plan_commercial_priority_rows(
    plan: PlanProduccion,
    *,
    lookback_days: int = 60,
    limit: int = 6,
) -> list[dict[str, object]]:
    plan_items = list(plan.items.select_related("receta").all())
    if not plan_items:
        return []

    plan_qty_map: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
    receta_ids: list[int] = []
    for item in plan_items:
        plan_qty_map[item.receta_id] += Decimal(str(item.cantidad or 0))
        receta_ids.append(item.receta_id)
    receta_ids = sorted(set(receta_ids))
    if not receta_ids:
        return []

    historico_map = {
        int(row["receta_id"]): Decimal(str(row["total"] or 0))
        for row in (
            VentaHistorica.objects.filter(
                receta_id__in=receta_ids,
                fecha__gte=plan.fecha_produccion - timedelta(days=lookback_days),
                fecha__lt=plan.fecha_produccion,
            )
            .values("receta_id")
            .annotate(total=Sum("cantidad"))
        )
    }
    solicitud_map = {
        int(row["insumo_id"]): Decimal(str(row["total"] or 0))
        for row in (
            SolicitudCompra.objects.filter(area=f"PLAN_PRODUCCION:{plan.id}", insumo_id__isnull=False)
            .values("insumo_id")
            .annotate(total=Sum("cantidad"))
        )
    }

    priority_map: dict[int, dict[str, object]] = {}
    for linea in (
        LineaReceta.objects.filter(receta_id__in=receta_ids, insumo_id__isnull=False)
        .exclude(tipo_linea=LineaReceta.TIPO_SUBSECCION)
        .select_related("insumo", "receta")
    ):
        if not linea.insumo_id or not linea.insumo:
            continue
        canonical = canonical_insumo(linea.insumo) or linea.insumo
        readiness_profile = enterprise_readiness_profile(canonical)
        row = priority_map.setdefault(
            linea.insumo_id,
            {
                "insumo_id": linea.insumo_id,
                "insumo_nombre": linea.insumo.nombre,
                "required_qty": Decimal("0"),
                "historico_units": Decimal("0"),
                "solicitud_qty": solicitud_map.get(linea.insumo_id, Decimal("0")),
                "recipe_names": set(),
                "priority_score": Decimal("0"),
                "master_missing": readiness_profile["missing"][:2] or ["Sin faltante"],
                "action_url": reverse("maestros:insumo_update", args=[canonical.id]),
            },
        )
        required_component = plan_qty_map.get(linea.receta_id, Decimal("0")) * Decimal(str(linea.cantidad or 0))
        historical_units = historico_map.get(linea.receta_id, Decimal("0"))
        row["required_qty"] += required_component
        row["historico_units"] += historical_units
        row["recipe_names"].add(linea.receta.nombre)
        row["priority_score"] += required_component * max(historical_units, Decimal("1"))

    rows: list[dict[str, object]] = []
    for row in priority_map.values():
        historico_units = Decimal(str(row["historico_units"] or 0))
        if historico_units >= Decimal("40"):
            priority_label = "Alta"
            priority_tone = "danger"
        elif historico_units >= Decimal("15"):
            priority_label = "Media"
            priority_tone = "warning"
        else:
            priority_label = "Base"
            priority_tone = "primary"
        rows.append(
            {
                **row,
                "recipe_names": sorted(row["recipe_names"])[:3],
                "priority_label": priority_label,
                "priority_tone": priority_tone,
            }
        )

    rows.sort(
        key=lambda item: (
            Decimal(str(item["priority_score"] or 0)),
            Decimal(str(item["historico_units"] or 0)),
            Decimal(str(item["required_qty"] or 0)),
        ),
        reverse=True,
    )
    return rows[:limit]


def _redirect_scoped_list(view_name: str, request: HttpRequest, *, preserve_query: bool = False) -> HttpResponse:
    if preserve_query:
        return_query = (request.POST.get("return_query") or "").strip()
        if return_query:
            return redirect(f"{reverse(view_name)}?{return_query}")

    source_filter = (request.POST.get("source") or "").strip().lower()
    plan_filter = (request.POST.get("plan_id") or "").strip()
    params: dict[str, str] = {}
    if source_filter in {"manual", "plan", "reabasto_cedis"}:
        params["source"] = source_filter
    if source_filter == "plan" and plan_filter:
        params["plan_id"] = plan_filter
    if params:
        return redirect(f"{reverse(view_name)}?{urlencode(params)}")
    return redirect(view_name)


def _locked_plan_scope_from_request(request: HttpRequest, *, current_view: str) -> dict | None:
    source_filter = (request.POST.get("source") or "").strip().lower()
    plan_filter = (request.POST.get("plan_id") or "").strip()
    if source_filter != "plan" or not plan_filter:
        return None
    return _build_plan_scope_context(
        source_filter="plan",
        plan_filter=plan_filter,
        current_view=current_view,
        session=request.session,
    )


def _scope_matches_plan_scope(*, plan_scope: str, area: str = "", referencia: str = "") -> bool:
    return (area or "").strip() == plan_scope or (referencia or "").strip() == plan_scope


def _source_context_from_scope(*, area: str = "", referencia: str = "", planes_map: dict[int, PlanProduccion] | None = None) -> dict[str, object]:
    planes_map = planes_map or {}
    source_scope = (referencia or area or "").strip()
    source_tipo = "manual"
    source_plan_id = None
    source_plan_nombre = ""
    source_label = "Manual"

    plan_id = _extract_plan_id_from_scope(source_scope)
    if plan_id:
        plan_id_int = int(plan_id)
        source_plan_id = plan_id_int
        plan_obj = planes_map.get(plan_id_int)
        source_plan_nombre = plan_obj.nombre if plan_obj else f"Plan {plan_id_int}"
        source_label = source_plan_nombre
        source_tipo = "plan"
        if plan_obj:
            plan_name = (plan_obj.nombre or "").strip().upper()
            plan_notes = (plan_obj.notas or "").strip().upper()
            if plan_name.startswith("CEDIS REABASTO ") or "[AUTO_REABASTO_CEDIS:" in plan_notes:
                source_tipo = "reabasto_cedis"
                source_label = "Reabasto CEDIS"

    return {
        "source_tipo": source_tipo,
        "source_plan_id": source_plan_id,
        "source_plan_nombre": source_plan_nombre,
        "source_label": source_label,
    }


def _build_import_preview_context(import_preview_payload) -> dict | None:
    if not isinstance(import_preview_payload, dict):
        return None

    preview_rows = [x for x in (import_preview_payload.get("rows") or []) if isinstance(x, dict)]
    try:
        preview_score_min = int(import_preview_payload.get("score_min") or 90)
    except (TypeError, ValueError):
        preview_score_min = 90

    preview_ready = 0
    preview_with_issues = 0
    preview_duplicates = 0
    preview_without_match = 0
    preview_invalid_qty = 0
    preview_ready_qty = Decimal("0")
    preview_ready_budget = Decimal("0")
    for row in preview_rows:
        row_include = bool(row.get("include"))
        if row_include:
            preview_ready += 1
        if row.get("notes"):
            preview_with_issues += 1
        if bool(row.get("duplicate")):
            preview_duplicates += 1
        if not str(row.get("insumo_id") or "").strip():
            preview_without_match += 1
        try:
            cantidad_preview = _to_decimal(str(row.get("cantidad") or "0"), "0")
        except Exception:
            cantidad_preview = Decimal("0")
        if cantidad_preview <= 0:
            preview_invalid_qty += 1
        if row_include and cantidad_preview > 0:
            preview_ready_qty += cantidad_preview
            presupuesto_estimado = _to_decimal(str(row.get("presupuesto_estimado") or "0"), "0")
            if presupuesto_estimado <= 0:
                costo_unitario = _to_decimal(str(row.get("costo_unitario") or "0"), "0")
                presupuesto_estimado = cantidad_preview * max(costo_unitario, Decimal("0"))
            preview_ready_budget += max(presupuesto_estimado, Decimal("0"))

    return {
        "rows": preview_rows,
        "count": len(preview_rows),
        "evitar_duplicados": bool(import_preview_payload.get("evitar_duplicados")),
        "score_min": preview_score_min,
        "ready_count": preview_ready,
        "excluded_count": max(0, len(preview_rows) - preview_ready),
        "issues_count": preview_with_issues,
        "duplicates_count": preview_duplicates,
        "without_match_count": preview_without_match,
        "invalid_qty_count": preview_invalid_qty,
        "ready_qty_total": preview_ready_qty,
        "ready_budget_total": preview_ready_budget,
        "file_name": str(import_preview_payload.get("file_name") or "").strip(),
        "generated_at": str(import_preview_payload.get("generated_at") or "").strip(),
    }


def _map_import_header(name: str) -> str:
    n = normalizar_nombre(name or "").replace("_", " ")
    if n in {"insumo", "nombre insumo", "insumo nombre", "materia prima", "articulo", "item", "producto", "descripcion"}:
        return "insumo"
    if n in {"cantidad", "cant", "qty", "cantidad requerida", "requerido"}:
        return "cantidad"
    if n in {"proveedor", "proveedor sugerido"}:
        return "proveedor"
    if n in {"fecha", "fecha requerida", "fecha requerida compra", "fecha requerida compras"}:
        return "fecha_requerida"
    if n in {"area", "area solicitante", "departamento"}:
        return "area"
    if n in {"solicitante", "responsable", "usuario"}:
        return "solicitante"
    if n in {"estatus", "estado"}:
        return "estatus"
    if n in {"periodo tipo", "tipo periodo", "tipo"}:
        return "periodo_tipo"
    if n in {"periodo mes", "mes", "periodo"}:
        return "periodo_mes"
    if n in {"monto objetivo", "presupuesto objetivo", "objetivo", "monto", "presupuesto"}:
        return "monto_objetivo"
    if n in {"monto objetivo proveedor", "objetivo proveedor", "presupuesto proveedor", "monto proveedor"}:
        return "monto_objetivo_proveedor"
    if n in {"categoria", "categoria insumo", "familia", "linea", "grupo"}:
        return "categoria"
    if n in {"monto objetivo categoria", "objetivo categoria", "presupuesto categoria", "monto categoria"}:
        return "monto_objetivo_categoria"
    if n in {"nota", "notas", "comentario", "comentarios"}:
        return "notas"
    return n


def _read_import_rows(uploaded) -> list[dict]:
    ext = Path(uploaded.name or "").suffix.lower()
    rows: list[dict] = []

    if ext in {".xlsx", ".xlsm"}:
        uploaded.seek(0)
        wb = load_workbook(uploaded, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        first_row = next(rows_iter, None)
        if not first_row:
            return []
        headers = [_map_import_header(str(h or "")) for h in first_row]
        for raw in rows_iter:
            row = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                row[header] = raw[idx] if idx < len(raw) else None
            rows.append(row)
        return rows

    if ext == ".csv":
        uploaded.seek(0)
        content = uploaded.read().decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(StringIO(content))
        for raw in reader:
            row = {}
            for k, v in raw.items():
                if not k:
                    continue
                row[_map_import_header(k)] = v
            rows.append(row)
        return rows

    raise ValueError("Formato no soportado. Usa .xlsx, .xlsm o .csv.")


@login_required
def descargar_plantilla_solicitudes(request: HttpRequest) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para descargar plantilla de importación.")

    export_format = (request.GET.get("format") or "xlsx").strip().lower()
    headers = ["insumo", "cantidad", "proveedor", "fecha_requerida", "area", "solicitante", "estatus"]
    sample_rows = [
        ["Harina Pastelera", "12.500", "Proveedor A", date.today().isoformat(), "Compras", request.user.username, SolicitudCompra.STATUS_BORRADOR],
        ["Mantequilla", "8.000", "Proveedor B", (date.today() + timedelta(days=2)).isoformat(), "Produccion", request.user.username, SolicitudCompra.STATUS_BORRADOR],
    ]

    if export_format == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="plantilla_solicitudes_compras.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(sample_rows)
        return response

    if export_format not in {"xlsx", "xlsm"}:
        messages.error(request, "Formato de plantilla no soportado. Usa csv o xlsx.")
        return redirect("compras:solicitudes")

    wb = Workbook()
    ws = wb.active
    ws.title = "solicitudes_import"
    ws.append(headers)
    for row in sample_rows:
        ws.append(row)
    for col in ("A", "B", "C", "D", "E", "F", "G"):
        ws.column_dimensions[col].width = 24

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_solicitudes_compras.xlsx"'
    return response


def _parse_periodo_tipo_value(raw) -> str | None:
    value = normalizar_nombre(str(raw or ""))
    if value in {"mes", "mensual"}:
        return "mes"
    if value in {"q1", "1ra quincena", "primera quincena", "quincena 1", "q 1"}:
        return "q1"
    if value in {"q2", "2da quincena", "segunda quincena", "quincena 2", "q 2"}:
        return "q2"
    return None


def _parse_periodo_mes_value(raw) -> str | None:
    if not raw:
        return None
    if isinstance(raw, date):
        return f"{raw.year:04d}-{raw.month:02d}"
    text = str(raw).strip()
    if not text:
        return None
    text = text.replace("/", "-")
    try:
        y, m = text.split("-")[:2]
        yi = int(y)
        mi = int(m)
        if 1 <= mi <= 12:
            return f"{yi:04d}-{mi:02d}"
    except Exception:
        pass
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            dt = datetime.strptime(text, fmt).date()
            return f"{dt.year:04d}-{dt.month:02d}"
        except ValueError:
            continue
    return None


def _default_fecha_requerida(periodo_tipo: str, periodo_mes: str) -> date:
    if periodo_tipo == "all":
        return timezone.localdate()
    year, month = periodo_mes.split("-")
    y = int(year)
    m = int(month)
    if periodo_tipo == "q1":
        return date(y, m, 15)
    if periodo_tipo == "q2":
        return date(y, m, calendar.monthrange(y, m)[1])
    return date(y, m, 1)


def _parse_date_value(raw_value, fallback: date) -> date:
    if not raw_value:
        return fallback
    if isinstance(raw_value, date):
        return raw_value
    text = str(raw_value).strip()
    if not text:
        return fallback
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return fallback


def _resolve_proveedor_name(raw: str, providers_by_norm: dict[str, Proveedor]) -> Proveedor | None:
    name = (raw or "").strip()
    if not name:
        return None
    return providers_by_norm.get(normalizar_nombre(name))


def _normalize_categoria_text(raw: str) -> str:
    return normalizar_nombre((raw or "").strip())


def _sanitize_categoria_filter(raw: str) -> str:
    return " ".join((raw or "").strip().split())


def _sanitize_consumo_ref_filter(raw: str) -> str:
    value = (raw or "all").strip().lower()
    if value not in {"all", "plan_ref"}:
        return "all"
    return value


def _resolve_insumo_categoria(insumo: Insumo) -> str:
    categoria = " ".join((getattr(insumo, "categoria", "") or "").strip().split())
    if categoria:
        return categoria
    unidad = getattr(insumo, "unidad_base", None)
    if not unidad:
        return "Sin categoría"
    tipo = (unidad.tipo or "").strip().upper()
    if tipo == "MASS":
        return "Masa"
    if tipo == "VOLUME":
        return "Volumen"
    if tipo == "UNIT":
        return "Pieza"
    return "Sin categoría"


def _write_import_pending_csv(rows: list[dict]) -> str:
    ts = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    filepath = Path("logs") / f"compras_import_pendientes_{ts}.csv"
    filepath.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "row",
        "insumo_origen",
        "cantidad_origen",
        "score",
        "metodo",
        "sugerencia",
        "motivo",
    ]
    with filepath.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({h: row.get(h, "") for h in headers})
    return str(filepath)


def _export_import_preview_csv(import_preview: dict) -> HttpResponse:
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="compras_preview_solicitudes_{now_str}.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "row_id",
            "source_row",
            "include",
            "insumo_origen",
            "insumo_sugerencia",
            "insumo_id",
            "cantidad",
            "area",
            "solicitante",
            "fecha_requerida",
            "estatus",
            "proveedor_id",
            "score",
            "metodo",
            "costo_unitario",
            "presupuesto_estimado",
            "duplicate",
            "notes",
        ]
    )
    for row in import_preview.get("rows", []):
        writer.writerow(
            [
                row.get("row_id", ""),
                row.get("source_row", ""),
                "1" if row.get("include") else "0",
                row.get("insumo_origen", ""),
                row.get("insumo_sugerencia", ""),
                row.get("insumo_id", ""),
                row.get("cantidad", ""),
                row.get("area", ""),
                row.get("solicitante", ""),
                row.get("fecha_requerida", ""),
                row.get("estatus", ""),
                row.get("proveedor_id", ""),
                row.get("score", ""),
                row.get("metodo", ""),
                row.get("costo_unitario", ""),
                row.get("presupuesto_estimado", ""),
                "1" if row.get("duplicate") else "0",
                row.get("notes", ""),
            ]
        )
    return response


def _export_import_preview_xlsx(import_preview: dict) -> HttpResponse:
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    wb = Workbook()
    ws = wb.active
    ws.title = "preview_import"
    ws.append(
        [
            "row_id",
            "source_row",
            "include",
            "insumo_origen",
            "insumo_sugerencia",
            "insumo_id",
            "cantidad",
            "area",
            "solicitante",
            "fecha_requerida",
            "estatus",
            "proveedor_id",
            "score",
            "metodo",
            "costo_unitario",
            "presupuesto_estimado",
            "duplicate",
            "notes",
        ]
    )
    for row in import_preview.get("rows", []):
        ws.append(
            [
                row.get("row_id", ""),
                row.get("source_row", ""),
                1 if row.get("include") else 0,
                row.get("insumo_origen", ""),
                row.get("insumo_sugerencia", ""),
                row.get("insumo_id", ""),
                row.get("cantidad", ""),
                row.get("area", ""),
                row.get("solicitante", ""),
                row.get("fecha_requerida", ""),
                row.get("estatus", ""),
                row.get("proveedor_id", ""),
                row.get("score", ""),
                row.get("metodo", ""),
                row.get("costo_unitario", ""),
                row.get("presupuesto_estimado", ""),
                1 if row.get("duplicate") else 0,
                row.get("notes", ""),
            ]
        )
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="compras_preview_solicitudes_{now_str}.xlsx"'
    return response


def _active_solicitud_statuses() -> set[str]:
    return {
        SolicitudCompra.STATUS_BORRADOR,
        SolicitudCompra.STATUS_EN_REVISION,
        SolicitudCompra.STATUS_APROBADA,
    }


def _can_transition_solicitud(current: str, new: str) -> bool:
    transitions = {
        SolicitudCompra.STATUS_BORRADOR: {SolicitudCompra.STATUS_EN_REVISION, SolicitudCompra.STATUS_APROBADA, SolicitudCompra.STATUS_RECHAZADA},
        SolicitudCompra.STATUS_EN_REVISION: {SolicitudCompra.STATUS_APROBADA, SolicitudCompra.STATUS_RECHAZADA},
        SolicitudCompra.STATUS_APROBADA: set(),
        SolicitudCompra.STATUS_RECHAZADA: set(),
    }
    return new in transitions.get(current, set())


def _can_transition_orden(current: str, new: str) -> bool:
    transitions = {
        OrdenCompra.STATUS_BORRADOR: {OrdenCompra.STATUS_ENVIADA},
        OrdenCompra.STATUS_ENVIADA: {OrdenCompra.STATUS_CONFIRMADA, OrdenCompra.STATUS_PARCIAL},
        OrdenCompra.STATUS_CONFIRMADA: {OrdenCompra.STATUS_PARCIAL, OrdenCompra.STATUS_CERRADA},
        OrdenCompra.STATUS_PARCIAL: {OrdenCompra.STATUS_CERRADA},
        OrdenCompra.STATUS_CERRADA: set(),
    }
    return new in transitions.get(current, set())


def _can_transition_recepcion(current: str, new: str) -> bool:
    transitions = {
        RecepcionCompra.STATUS_PENDIENTE: {RecepcionCompra.STATUS_DIFERENCIAS, RecepcionCompra.STATUS_CERRADA},
        RecepcionCompra.STATUS_DIFERENCIAS: {RecepcionCompra.STATUS_CERRADA},
        RecepcionCompra.STATUS_CERRADA: set(),
    }
    return new in transitions.get(current, set())


def _build_status_cards(items, definitions, *, get_status=None):
    status_getter = get_status or (lambda item: getattr(item, "estatus", ""))
    cards = []
    for code, label, tone in definitions:
        count = sum(1 for item in items if status_getter(item) == code)
        cards.append(
            {
                "code": code,
                "label": label,
                "tone": tone,
                "count": count,
            }
        )
    return cards


def _enrich_solicitud_workflow(solicitud):
    solicitud.workflow_blocker_details = _enterprise_blocker_details_for_solicitud(solicitud)
    solicitud.workflow_blockers = [item["message"] for item in solicitud.workflow_blocker_details]
    solicitud.has_workflow_blockers = bool(solicitud.workflow_blockers)
    solicitud.enterprise_master_summary = _enterprise_master_summary_from_details(solicitud.workflow_blocker_details)
    if solicitud.estatus == SolicitudCompra.STATUS_BORRADOR:
        solicitud.workflow_stage = "Captura"
        solicitud.workflow_next = "Corregir bloqueos ERP" if solicitud.has_workflow_blockers else "Enviar a revisión"
    elif solicitud.estatus == SolicitudCompra.STATUS_EN_REVISION:
        solicitud.workflow_stage = "Validación"
        solicitud.workflow_next = "Corregir bloqueos ERP" if solicitud.has_workflow_blockers else "Aprobar o rechazar"
    elif solicitud.estatus == SolicitudCompra.STATUS_APROBADA and solicitud.has_open_order:
        solicitud.workflow_stage = "Abastecimiento"
        solicitud.workflow_next = f"Dar seguimiento a OC {solicitud.open_order_folio}"
    elif solicitud.estatus == SolicitudCompra.STATUS_APROBADA:
        solicitud.workflow_stage = "Lista para compra"
        solicitud.workflow_next = "Crear orden de compra"
    else:
        solicitud.workflow_stage = "Cerrada"
        solicitud.workflow_next = "Sin acción"
    if solicitud.has_workflow_blockers:
        solicitud.workflow_health_label = "Bloqueada ERP"
        solicitud.workflow_health_tone = "danger"
    elif solicitud.estatus == SolicitudCompra.STATUS_APROBADA and solicitud.has_open_order:
        solicitud.workflow_health_label = "Abastecimiento"
        solicitud.workflow_health_tone = "primary"
    elif solicitud.estatus == SolicitudCompra.STATUS_APROBADA:
        solicitud.workflow_health_label = "Lista para compra"
        solicitud.workflow_health_tone = "success"
    elif solicitud.estatus == SolicitudCompra.STATUS_EN_REVISION:
        solicitud.workflow_health_label = "En validación"
        solicitud.workflow_health_tone = "warning"
    elif solicitud.estatus == SolicitudCompra.STATUS_RECHAZADA:
        solicitud.workflow_health_label = "Cerrada"
        solicitud.workflow_health_tone = "muted"
    else:
        solicitud.workflow_health_label = "En captura"
        solicitud.workflow_health_tone = "warning"

    if solicitud.has_workflow_blockers:
        solicitud.workflow_action_label = "Corregir maestro ERP"
        solicitud.workflow_action_code = "corregir_maestro"
    elif solicitud.estatus == SolicitudCompra.STATUS_BORRADOR:
        solicitud.workflow_action_label = "Enviar a revisión"
        solicitud.workflow_action_code = "enviar_revision"
    elif solicitud.estatus == SolicitudCompra.STATUS_EN_REVISION:
        solicitud.workflow_action_label = "Aprobar o rechazar"
        solicitud.workflow_action_code = "aprobar_rechazar"
    elif solicitud.estatus == SolicitudCompra.STATUS_APROBADA and solicitud.has_open_order:
        solicitud.workflow_action_label = "Dar seguimiento a OC"
        solicitud.workflow_action_code = "seguimiento_oc"
    elif solicitud.estatus == SolicitudCompra.STATUS_APROBADA:
        solicitud.workflow_action_label = "Crear OC"
        solicitud.workflow_action_code = "crear_oc"
    else:
        solicitud.workflow_action_label = "Sin acción"
        solicitud.workflow_action_code = "none"
    return solicitud


def _enterprise_blocker_action_meta(
    insumo: Insumo | None,
    missing_field: str,
    *,
    usage_scope: str = "purchases",
) -> dict[str, str]:
    normalized = (missing_field or "").strip().lower()
    if not insumo:
        return {
            "label": "Revisar catálogo",
            "url": reverse("maestros:insumo_list"),
            "detail": "Confirma el artículo canónico antes de continuar el flujo documental.",
        }

    action_map = {
        "unidad base": (
            "Definir unidad base",
            "unidad",
            "Completa la unidad base del artículo para permitir compras, costeo e inventario.",
        ),
        "proveedor principal": (
            "Asignar proveedor principal",
            "proveedor",
            "Asigna el proveedor principal para que compras pueda sugerir abastecimiento correcto.",
        ),
        "categoría": (
            "Asignar categoría",
            "categoria",
            "Completa la categoría operativa para clasificar el artículo en compras y reportes.",
        ),
        "código point": (
            "Registrar código comercial",
            "codigo_point",
            "Captura el código comercial para mantener catálogo y reportes consistentes.",
        ),
        "codigo point": (
            "Registrar código comercial",
            "codigo_point",
            "Captura el código comercial para mantener catálogo y reportes consistentes.",
        ),
        "costo vigente": (
            "Cargar costo vigente",
            "",
            "Registra un costo vigente para que el documento pueda avanzar sin bloquear presupuesto.",
        ),
        "inactivo": (
            "Reactivar artículo",
            "",
            "Activa nuevamente el artículo o sustituye el documento por un artículo vigente.",
        ),
        "catálogo canónico": (
            "Depurar artículo maestro",
            "",
            "Reapunta el documento al artículo canónico para evitar duplicados operativos.",
        ),
        "catalogo canonico": (
            "Depurar artículo maestro",
            "",
            "Reapunta el documento al artículo canónico para evitar duplicados operativos.",
        ),
    }
    action_label, query_key, detail = action_map.get(
        normalized,
        (
            "Abrir artículo",
            "",
            "Revisa el artículo en maestro para completar la información faltante.",
        ),
    )
    if query_key:
        url = (
            f"{reverse('maestros:insumo_list')}?usage_scope={usage_scope}"
            f"&enterprise_status=incompletos&missing_field={query_key}&q={insumo.nombre}"
        )
    else:
        url = reverse("maestros:insumo_update", args=[insumo.id])
    return {"label": action_label, "url": url, "detail": detail}


def _enterprise_blocker_label_detail_for_missing(missing_field: str | None) -> tuple[str, str]:
    normalized = (missing_field or "").strip().lower()
    if normalized == "unidad base":
        return "Definir unidad base", "Revisa los artículos de esta clase sin unidad base para desbloquear compras."
    if normalized == "proveedor principal":
        return "Asignar proveedor principal", "Revisa los artículos de esta clase sin proveedor principal para habilitar abastecimiento."
    if normalized in {"categoría", "categoria"}:
        return "Asignar categoría", "Completa la categoría operativa para ordenar compras y reportes."
    if normalized in {"código point", "codigo point"}:
        return "Registrar código comercial", "Captura el código comercial para mantener catálogo y reportes consistentes."
    if normalized in {"catálogo canónico", "catalogo canonico"}:
        return "Depurar artículo maestro", "Reapunta el documento al artículo maestro para evitar duplicados operativos."
    if normalized == "costo vigente":
        return "Cargar costo vigente", "Registra el costo vigente para desbloquear presupuesto y órdenes."
    if normalized == "inactivo":
        return "Reactivar artículo", "Activa o sustituye los artículos bloqueados de esta clase."
    return "Abrir maestro", "Revisa la clase con mayor bloqueo del maestro."


def _enterprise_blocker_action_for_missing(insumo: Insumo | None, missing_field: str, *, usage_scope: str = "purchases") -> tuple[str, str]:
    meta = _enterprise_blocker_action_meta(insumo, missing_field, usage_scope=usage_scope)
    return meta["label"], meta["url"]


def _enterprise_article_class(insumo: Insumo | None) -> dict[str, str]:
    if not insumo:
        return {"key": "unknown", "label": "Sin catálogo"}
    if insumo.tipo_item == Insumo.TIPO_EMPAQUE:
        return {"key": Insumo.TIPO_EMPAQUE, "label": "Empaque"}
    if insumo.tipo_item == Insumo.TIPO_INTERNO or (insumo.codigo or "").startswith("DERIVADO:RECETA:"):
        return {"key": Insumo.TIPO_INTERNO, "label": "Insumo interno"}
    return {"key": Insumo.TIPO_MATERIA_PRIMA, "label": "Materia prima"}


def _missing_field_filter_key(missing_field: str | None) -> str | None:
    mapping = {
        "unidad base": "unidad",
        "proveedor principal": "proveedor",
        "categoría": "categoria",
        "código point": "codigo_point",
        "codigo point": "codigo_point",
        "catálogo canónico": "canonico",
        "catalogo canonico": "canonico",
    }
    return mapping.get((missing_field or "").strip().lower())


def _enterprise_blocker_details_for_solicitud(solicitud) -> list[dict[str, object]]:
    original_insumo = getattr(solicitud, "insumo", None)
    if original_insumo is None and getattr(solicitud, "insumo_id", None):
        try:
            original_insumo = Insumo.objects.filter(id=solicitud.insumo_id).first()
        except Exception:
            original_insumo = None
    canonical = canonical_insumo_by_id(solicitud.insumo_id) if getattr(solicitud, "insumo_id", None) else None
    insumo = canonical or original_insumo
    if not insumo:
        action_meta = _enterprise_blocker_action_meta(None, "catalogo")
        return [
            {
                "key": "sin_catalogo",
                "message": "Artículo no encontrado en catálogo canónico",
                "insumo_id": None,
                "insumo_nombre": getattr(solicitud, "insumo_texto", "Artículo sin catálogo") or "Artículo sin catálogo",
                "missing_field": "catálogo canónico",
                "action_label": action_meta["label"],
                "action_url": action_meta["url"],
                "action_detail": action_meta["detail"],
                "tone": "danger",
            }
        ]

    profile = enterprise_readiness_profile(insumo)
    blockers: list[dict[str, object]] = []
    if original_insumo and canonical and original_insumo.id != canonical.id:
        action_meta = _enterprise_blocker_action_meta(canonical, "catálogo canónico")
        blockers.append(
            {
                "key": "no_canonico",
                "message": f"Ligado a variante no canónica: usar {canonical.nombre}",
                "insumo_id": canonical.id,
                "insumo_nombre": canonical.nombre,
                "missing_field": "catálogo canónico",
                "action_label": action_meta["label"],
                "action_url": action_meta["url"],
                "action_detail": action_meta["detail"],
                "tone": "warning",
            }
        )
    if profile["readiness_label"] == "Inactivo":
        action_meta = _enterprise_blocker_action_meta(insumo, "inactivo")
        blockers.append(
            {
                "key": "articulo_inactivo",
                "message": "Artículo inactivo en maestro",
                "insumo_id": insumo.id,
                "insumo_nombre": insumo.nombre,
                "missing_field": "inactivo",
                "action_label": action_meta["label"],
                "action_url": action_meta["url"],
                "action_detail": action_meta["detail"],
                "tone": "danger",
            }
        )
    elif profile["readiness_label"] == "Incompleto":
        blockers.append(
            {
                "key": "maestro_incompleto",
                "message": "Artículo incompleto: " + ", ".join(profile["missing"]),
                "insumo_id": insumo.id,
                "insumo_nombre": insumo.nombre,
                "missing_field": ", ".join(profile["missing"]),
                "action_label": "Corregir artículo",
                "action_url": reverse("maestros:insumo_update", args=[insumo.id]),
                "action_detail": "Completa los faltantes críticos del artículo para desbloquear compras.",
                "tone": "warning",
            }
        )
        for missing in profile["missing"]:
            action_meta = _enterprise_blocker_action_meta(insumo, missing)
            blockers.append(
                {
                    "key": f"missing_{normalizar_nombre(missing).replace(' ', '_')}",
                    "message": f"Falta {missing}",
                    "insumo_id": insumo.id,
                    "insumo_nombre": insumo.nombre,
                    "missing_field": missing,
                    "action_label": action_meta["label"],
                    "action_url": action_meta["url"],
                    "action_detail": action_meta["detail"],
                    "tone": "warning",
                }
            )

    latest_cost = latest_costo_canonico(insumo_id=insumo.id)
    if latest_cost is None or latest_cost <= 0:
        blockers.append(
            {
                "key": "sin_costo",
                "message": "Sin costo vigente",
                "insumo_id": insumo.id,
                "insumo_nombre": insumo.nombre,
                "missing_field": "costo vigente",
                "action_label": "Cargar costo vigente",
                "action_url": reverse("maestros:insumo_update", args=[insumo.id]),
                "action_detail": "Registra un costo vigente para habilitar presupuesto, OC y recepción.",
                "tone": "danger",
            }
        )

    proveedor = solicitud.proveedor_sugerido or insumo.proveedor_principal
    if not proveedor:
        action_meta = _enterprise_blocker_action_meta(insumo, "proveedor principal")
        blockers.append(
            {
                "key": "sin_proveedor",
                "message": "Sin proveedor sugerido",
                "insumo_id": insumo.id,
                "insumo_nombre": insumo.nombre,
                "missing_field": "proveedor principal",
                "action_label": action_meta["label"],
                "action_url": action_meta["url"],
                "action_detail": action_meta["detail"],
                "tone": "danger",
            }
        )
    return blockers


def _enterprise_blockers_for_solicitud(solicitud) -> list[str]:
    return [item["message"] for item in _enterprise_blocker_details_for_solicitud(solicitud)]


def _enterprise_master_summary_from_details(detail_rows: list[dict[str, object]] | None) -> dict[str, object]:
    rows = list(detail_rows or [])
    master_rows = [
        row
        for row in rows
        if row.get("missing_field")
        and str(row.get("missing_field") or "").strip().lower() != "costo vigente"
    ]
    if not master_rows:
        return {
            "status_label": "Listo ERP",
            "status_tone": "success",
            "missing_labels": [],
            "action_label": "",
            "action_url": "",
            "action_detail": "",
            "summary": "Artículo completo para operar en compras.",
        }
    first = master_rows[0]
    missing_labels: list[str] = []
    for row in master_rows:
        label = str(row.get("missing_field") or "").strip()
        if label and label not in missing_labels:
            missing_labels.append(label)
    return {
        "status_label": "Incompleto",
        "status_tone": "danger",
        "missing_labels": missing_labels,
        "action_label": first.get("action_label") or "Abrir maestro",
        "action_url": first.get("action_url") or reverse("maestros:insumo_list"),
        "action_detail": first.get("action_detail") or "Completa los faltantes del artículo para desbloquear compras.",
        "summary": "Falta " + ", ".join(missing_labels),
    }


def _enterprise_master_blocker_rollup(document_rows, *, usage_scope: str = "purchases") -> dict[str, list[dict[str, object]]]:
    class_registry: dict[str, dict[str, object]] = {}
    detail_registry: dict[tuple[str, str, str], dict[str, object]] = {}

    for row in document_rows:
        folio = row.get("folio") or "Sin folio"
        for detail in row.get("details") or []:
            insumo_id = detail.get("insumo_id")
            insumo = canonical_insumo_by_id(int(insumo_id)) if insumo_id else None
            article_class = _enterprise_article_class(insumo)
            missing_label = str(detail.get("missing_field") or "sin catálogo")

            class_bucket = class_registry.setdefault(
                article_class["key"],
                {
                    "class_key": article_class["key"],
                    "class_label": article_class["label"],
                    "count": 0,
                    "missing_counts": {},
                },
            )
            class_bucket["count"] += 1
            class_bucket["missing_counts"][missing_label] = class_bucket["missing_counts"].get(missing_label, 0) + 1

            detail_key = (
                article_class["key"],
                str(detail.get("insumo_nombre") or "Artículo sin catálogo"),
                missing_label,
            )
            detail_row = detail_registry.setdefault(
                detail_key,
                {
                    "class_key": article_class["key"],
                    "class_label": article_class["label"],
                    "insumo_id": int(insumo_id) if insumo_id else None,
                    "insumo_nombre": detail.get("insumo_nombre") or "Artículo sin catálogo",
                    "missing_field": missing_label,
                    "message": detail.get("message"),
                    "count": 0,
                    "folios": [],
                    "tone": detail.get("tone") or "warning",
                    "action_label": detail.get("action_label") or "Abrir maestro",
                    "action_url": detail.get("action_url") or reverse("maestros:insumo_list"),
                    "action_detail": detail.get("action_detail") or "Corrige el faltante maestro antes de continuar.",
                    "edit_url": "",
                },
            )
            detail_row["count"] += 1
            detail_row["folios"].append(folio)

            if insumo:
                filter_key = _missing_field_filter_key(missing_label)
                action_meta = _enterprise_blocker_action_meta(insumo, missing_label, usage_scope=usage_scope)
                action_url = (
                    f"{reverse('maestros:insumo_list')}?tipo_item={article_class['key']}&enterprise_status=incompletos"
                    f"&usage_scope={usage_scope}&insumo_id={insumo.id}"
                )
                if filter_key:
                    action_url += f"&missing_field={filter_key}"
                detail_row["action_label"] = action_meta["label"]
                detail_row["action_url"] = action_url
                detail_row["action_detail"] = action_meta["detail"]
                detail_row["edit_url"] = reverse("maestros:insumo_update", args=[insumo.id])

    class_cards = []
    missing_registry: dict[str, dict[str, object]] = {}
    for bucket in class_registry.values():
        missing_counts = bucket.pop("missing_counts")
        dominant_label = ""
        dominant_count = 0
        if missing_counts:
            dominant_label, dominant_count = max(missing_counts.items(), key=lambda item: (item[1], item[0]))
        filter_key = _missing_field_filter_key(dominant_label)
        action_label, action_detail = _enterprise_blocker_label_detail_for_missing(dominant_label)
        action_url = (
            f"{reverse('maestros:insumo_list')}?tipo_item={bucket['class_key']}&enterprise_status=incompletos&usage_scope={usage_scope}"
        )
        if filter_key:
            action_url += f"&missing_field={filter_key}"
        class_cards.append(
            {
                **bucket,
                "dominant_label": dominant_label or "sin clasificar",
                "dominant_count": dominant_count,
                "action_label": action_label if dominant_label else "Abrir maestro",
                "action_url": action_url,
                "action_detail": action_detail if dominant_label else "Revisa la clase con mayor bloqueo del maestro.",
            }
        )
        for missing_label, count in missing_counts.items():
            missing_key = _missing_field_filter_key(missing_label) or "other"
            missing_entry = missing_registry.setdefault(
                missing_key,
                {
                    "key": missing_key,
                    "missing_label": missing_label,
                    "count": 0,
                    "class_counts": {},
                },
            )
            missing_entry["count"] += int(count)
            missing_entry["class_counts"][bucket["class_label"]] = (
                missing_entry["class_counts"].get(bucket["class_label"], 0) + int(count)
            )

    missing_cards = []
    for item in missing_registry.values():
        class_counts = item.pop("class_counts")
        dominant_class_label = ""
        dominant_class_count = 0
        if class_counts:
            dominant_class_label, dominant_class_count = max(class_counts.items(), key=lambda value: (value[1], value[0]))
        action_label, action_detail = _enterprise_blocker_label_detail_for_missing(item["missing_label"])
        action_url = f"{reverse('maestros:insumo_list')}?enterprise_status=incompletos&usage_scope={usage_scope}"
        if item["key"] != "other":
            action_url += f"&missing_field={item['key']}"
        missing_cards.append(
            {
                **item,
                "dominant_class_label": dominant_class_label or "Sin clasificar",
                "dominant_class_count": dominant_class_count,
                "action_label": action_label,
                "action_url": action_url,
                "action_detail": action_detail,
            }
        )

    class_cards = sorted(class_cards, key=lambda item: (-int(item["count"]), str(item["class_label"])))
    missing_cards = sorted(missing_cards, key=lambda item: (-int(item["count"]), str(item["missing_label"])))
    detail_rows = sorted(
        detail_registry.values(),
        key=lambda row: (-int(row["count"]), str(row["class_label"]), str(row["insumo_nombre"]), str(row["missing_field"])),
    )[:12]
    for row in detail_rows:
        row["folios_preview"] = ", ".join(row["folios"][:3])
        row["folios_more"] = max(len(row["folios"]) - 3, 0)

    return {
        "master_blocker_class_cards": class_cards,
        "master_blocker_missing_cards": missing_cards,
        "master_blocker_detail_rows": detail_rows,
    }


def _solicitudes_workflow_summary(solicitudes):
    blocked_erp = sum(1 for item in solicitudes if getattr(item, "has_workflow_blockers", False))
    approved_ready = sum(
        1
        for item in solicitudes
        if item.estatus == SolicitudCompra.STATUS_APROBADA and not item.has_open_order
    )
    approved_with_oc = sum(
        1
        for item in solicitudes
        if item.estatus == SolicitudCompra.STATUS_APROBADA and item.has_open_order
    )
    cards = _build_status_cards(
        solicitudes,
        [
            (SolicitudCompra.STATUS_BORRADOR, "Borrador", "warning"),
            (SolicitudCompra.STATUS_EN_REVISION, "En revisión", "warning"),
            (SolicitudCompra.STATUS_RECHAZADA, "Rechazadas", "danger"),
        ],
    )
    cards.insert(
        2,
        {
            "code": "BLOCKED_ERP",
            "label": "Bloqueadas ERP",
            "tone": "danger",
            "count": blocked_erp,
            "query": "?estatus=BLOCKED_ERP",
        },
    )
    cards.insert(
        3,
        {
            "code": "APPROVED_READY",
            "label": "Listas para OC",
            "tone": "success",
            "count": approved_ready,
            "query": "?estatus=APPROVED_READY",
        },
    )
    cards.insert(
        4,
        {
            "code": "APPROVED_WITH_OC",
            "label": "Con OC activa",
            "tone": "primary",
            "count": approved_with_oc,
            "query": "?estatus=APPROVED_WITH_OC",
        },
    )
    for card in cards:
        card.setdefault("query", f"?estatus={card['code']}")
    total = len(solicitudes)
    gate_cards = [
        {
            "key": "master_ready",
            "label": "Maestro listo",
            "ready_count": max(total - blocked_erp, 0),
            "blocked_count": blocked_erp,
            "tone": "success" if blocked_erp == 0 else "danger",
            "detail": "Solicitudes sin bloqueo ERP y con datos maestros suficientes para seguir el flujo.",
            "action_label": "Resolver bloqueos",
            "query": "?estatus=BLOCKED_ERP",
        },
        {
            "key": "review_complete",
            "label": "Validación resuelta",
            "ready_count": sum(
                1
                for item in solicitudes
                if item.estatus in {SolicitudCompra.STATUS_APROBADA, SolicitudCompra.STATUS_RECHAZADA}
            ),
            "blocked_count": sum(1 for item in solicitudes if item.estatus in {SolicitudCompra.STATUS_BORRADOR, SolicitudCompra.STATUS_EN_REVISION}),
            "tone": "success"
            if not any(item.estatus in {SolicitudCompra.STATUS_BORRADOR, SolicitudCompra.STATUS_EN_REVISION} for item in solicitudes)
            else "warning",
            "detail": "La solicitud ya fue revisada y no sigue detenida en captura o revisión.",
            "action_label": "Abrir revisión",
            "query": "?workflow_action=aprobar_rechazar",
        },
        {
            "key": "ready_for_oc",
            "label": "Lista para OC",
            "ready_count": approved_ready,
            "blocked_count": sum(
                1
                for item in solicitudes
                if item.estatus == SolicitudCompra.STATUS_APROBADA and getattr(item, "has_open_order", False)
            ),
            "tone": "success" if approved_ready else "warning",
            "detail": "Solicitudes aprobadas y sin orden abierta, listas para convertirse en orden de compra.",
            "action_label": "Abrir listas",
            "query": "?estatus=APPROVED_READY",
        },
    ]
    return {
        "cards": cards,
        "gate_cards": gate_cards,
        "note": "Flujo documental: Captura -> Validación -> OC -> Recepción.",
    }


def _solicitudes_enterprise_board(solicitudes):
    next_step_defs = [
        ("corregir_maestro", "Corregir maestro ERP", "Corrección ERP", "danger"),
        ("enviar_revision", "Enviar a revisión", "Enviar revisión", "warning"),
        ("aprobar_rechazar", "Aprobar o rechazar", "Resolver validación", "warning"),
        ("crear_oc", "Crear OC", "Crear OC", "success"),
        ("seguimiento_oc", "Dar seguimiento a OC", "Seguimiento OC", "primary"),
    ]
    next_step_cards = []
    for key, action_label, label, tone in next_step_defs:
        next_step_cards.append(
            {
                "key": key,
                "label": label,
                "count": sum(1 for item in solicitudes if getattr(item, "workflow_action_label", "") == action_label),
                "tone": tone,
                "query": f"?workflow_action={key}",
                "action_detail": f"Abrir solicitudes en etapa {label.lower()}.",
            }
        )

    blocker_defs = [
        ("sin_costo", "Sin costo vigente", "danger"),
        ("sin_proveedor", "Sin proveedor", "danger"),
        ("maestro_incompleto", "Maestro incompleto", "warning"),
        ("articulo_inactivo", "Artículo inactivo", "danger"),
        ("sin_catalogo", "Sin catálogo canónico", "danger"),
    ]
    blocker_counts = {key: 0 for key, _, _ in blocker_defs}
    for item in solicitudes:
        for blocker in getattr(item, "workflow_blockers", []) or []:
            blocker_text = blocker.lower()
            if "sin costo vigente" in blocker_text:
                blocker_counts["sin_costo"] += 1
            elif "sin proveedor sugerido" in blocker_text:
                blocker_counts["sin_proveedor"] += 1
            elif "artículo incompleto" in blocker_text or "articulo incompleto" in blocker_text:
                blocker_counts["maestro_incompleto"] += 1
            elif "artículo inactivo" in blocker_text or "articulo inactivo" in blocker_text:
                blocker_counts["articulo_inactivo"] += 1
            elif "catálogo canónico" in blocker_text or "catalogo canonico" in blocker_text:
                blocker_counts["sin_catalogo"] += 1

    blocker_cards = [
        {
            "key": key,
            "label": label,
            "count": blocker_counts[key],
            "tone": tone,
            "query": f"?blocker_key={key}",
            "action_detail": f"Abrir solicitudes con bloqueo {label.lower()}.",
        }
        for key, label, tone in blocker_defs
    ]
    blocker_registry: dict[tuple[object, str], dict[str, object]] = {}
    for item in solicitudes:
        if not getattr(item, "has_workflow_blockers", False):
            continue
        for detail in getattr(item, "workflow_blocker_details", []) or []:
            registry_key = (detail.get("insumo_id") or detail.get("insumo_nombre"), detail.get("key"))
            row = blocker_registry.setdefault(
                registry_key,
                {
                    "insumo_id": detail.get("insumo_id"),
                    "insumo_nombre": detail.get("insumo_nombre"),
                    "missing_field": detail.get("missing_field"),
                    "message": detail.get("message"),
                    "action_label": detail.get("action_label"),
                    "action_url": detail.get("action_url"),
                    "action_detail": detail.get("action_detail") or "Corrige el faltante maestro antes de continuar.",
                    "tone": detail.get("tone"),
                    "count": 0,
                    "folios": [],
                },
            )
            row["count"] += 1
            row["folios"].append(item.folio)

    blocker_detail_rows = sorted(
        blocker_registry.values(),
        key=lambda row: (-int(row["count"]), str(row["insumo_nombre"] or ""), str(row["missing_field"] or "")),
    )[:10]
    for row in blocker_detail_rows:
        row["folios_preview"] = ", ".join(row["folios"][:3])
        row["folios_more"] = max(len(row["folios"]) - 3, 0)

    insumo_ids = {
        int(row["insumo_id"])
        for row in blocker_registry.values()
        if row.get("insumo_id")
    }
    canonical_map = {insumo_id: canonical_insumo_by_id(insumo_id) for insumo_id in insumo_ids}
    class_registry: dict[str, dict[str, object]] = {}
    missing_registry: dict[str, dict[str, object]] = {}
    master_blocker_detail_registry: dict[tuple[str, str, str], dict[str, object]] = {}
    for row in blocker_registry.values():
        insumo_id = row.get("insumo_id")
        insumo = canonical_map.get(int(insumo_id)) if insumo_id else None
        article_class = _enterprise_article_class(insumo)
        class_bucket = class_registry.setdefault(
            article_class["key"],
            {
                "class_key": article_class["key"],
                "class_label": article_class["label"],
                "count": 0,
                "missing_counts": {},
            },
        )
        class_bucket["count"] += int(row["count"])
        missing_label = str(row.get("missing_field") or "sin catálogo")
        class_bucket["missing_counts"][missing_label] = class_bucket["missing_counts"].get(missing_label, 0) + int(row["count"])

        missing_entry = missing_registry.setdefault(
            missing_label,
            {
                "key": _missing_field_filter_key(missing_label) or slugify(missing_label) or "missing",
                "missing_label": missing_label,
                "count": 0,
                "dominant_class_label": "",
                "dominant_class_count": 0,
                "class_counts": {},
                "action_label": "",
                "action_detail": "",
            },
        )
        missing_entry["count"] += int(row["count"])
        class_counts = missing_entry["class_counts"]
        class_counts[article_class["label"]] = class_counts.get(article_class["label"], 0) + int(row["count"])

        detail_key = (
            article_class["key"],
            str(row.get("insumo_nombre") or "Artículo sin catálogo"),
            missing_label,
        )
        detail_row = master_blocker_detail_registry.setdefault(
            detail_key,
            {
                "class_label": article_class["label"],
                "insumo_id": int(insumo_id) if insumo_id else None,
                "insumo_nombre": row.get("insumo_nombre") or "Artículo sin catálogo",
                "missing_field": missing_label,
                "message": row.get("message"),
                "count": 0,
                "folios": [],
                "tone": row.get("tone") or "warning",
                "action_label": row.get("action_label") or "Abrir maestro",
                "action_url": row.get("action_url") or reverse("maestros:insumo_list"),
                "action_detail": row.get("action_detail") or "Corrige el faltante maestro antes de continuar.",
                "edit_url": "",
            },
        )
        detail_row["count"] += int(row["count"])
        detail_row["folios"].extend(list(row.get("folios") or []))

        if insumo:
            filter_key = _missing_field_filter_key(missing_label)
            action_meta = _enterprise_blocker_action_meta(insumo, missing_label, usage_scope="purchases")
            action_url = (
                f"{reverse('maestros:insumo_list')}?tipo_item={article_class['key']}&enterprise_status=incompletos"
                f"&usage_scope=purchases&insumo_id={insumo.id}"
            )
            if filter_key:
                action_url += f"&missing_field={filter_key}"
            detail_row["action_label"] = action_meta["label"]
            detail_row["action_url"] = action_url
            detail_row["action_detail"] = action_meta["detail"]
            detail_row["edit_url"] = reverse("maestros:insumo_update", args=[insumo.id])

    master_blocker_class_cards = []
    for bucket in class_registry.values():
        missing_counts = bucket.pop("missing_counts")
        dominant_label = ""
        dominant_count = 0
        if missing_counts:
            dominant_label, dominant_count = max(missing_counts.items(), key=lambda item: (item[1], item[0]))
        filter_key = _missing_field_filter_key(dominant_label)
        action_label, action_detail = _enterprise_blocker_label_detail_for_missing(dominant_label)
        action_url = (
            f"{reverse('maestros:insumo_list')}?tipo_item={bucket['class_key']}&enterprise_status=incompletos&usage_scope=purchases"
        )
        if filter_key:
            action_url += f"&missing_field={filter_key}"
        master_blocker_class_cards.append(
            {
                **bucket,
                "dominant_label": dominant_label or "sin clasificar",
                "dominant_count": dominant_count,
                "action_label": action_label if dominant_label else "Abrir maestro",
                "action_url": action_url,
                "action_detail": action_detail if dominant_label else "Revisa la clase con mayor bloqueo del maestro.",
            }
        )

    master_blocker_class_cards = sorted(
        master_blocker_class_cards,
        key=lambda item: (-int(item["count"]), str(item["class_label"])),
    )
    master_blocker_missing_cards = []
    for item in missing_registry.values():
        class_counts = item.pop("class_counts")
        if class_counts:
            dominant_class_label, dominant_class_count = max(class_counts.items(), key=lambda pair: (pair[1], pair[0]))
        else:
            dominant_class_label, dominant_class_count = ("", 0)
        action_label, action_detail = _enterprise_blocker_label_detail_for_missing(item["missing_label"])
        item["dominant_class_label"] = dominant_class_label
        item["dominant_class_count"] = dominant_class_count
        item["action_label"] = action_label
        item["action_detail"] = action_detail
        master_blocker_missing_cards.append(item)
    master_blocker_missing_cards = sorted(
        master_blocker_missing_cards,
        key=lambda item: (-int(item["count"]), str(item["missing_label"])),
    )
    master_blocker_detail_rows = sorted(
        master_blocker_detail_registry.values(),
        key=lambda row: (-int(row["count"]), str(row["class_label"]), str(row["insumo_nombre"]), str(row["missing_field"])),
    )[:12]
    for row in master_blocker_detail_rows:
        row["folios_preview"] = ", ".join(row["folios"][:3])
        row["folios_more"] = max(len(row["folios"]) - 3, 0)

    return {
        "next_step_cards": next_step_cards,
        "blocker_cards": blocker_cards,
        "blocker_detail_rows": blocker_detail_rows,
        "master_blocker_class_cards": master_blocker_class_cards,
        "master_blocker_missing_cards": master_blocker_missing_cards,
        "master_blocker_detail_rows": master_blocker_detail_rows,
        "blocked_total": sum(1 for item in solicitudes if getattr(item, "has_workflow_blockers", False)),
        "ready_for_oc_total": sum(
            1
            for item in solicitudes
            if item.estatus == SolicitudCompra.STATUS_APROBADA and not getattr(item, "has_open_order", False) and not getattr(item, "has_workflow_blockers", False)
        ),
    }


def _solicitudes_supply_model_rows(solicitudes) -> list[dict[str, object]]:
    total = len(solicitudes)
    maestro_ready = sum(1 for item in solicitudes if not getattr(item, "has_workflow_blockers", False))
    validacion_ready = sum(
        1
        for item in solicitudes
        if item.estatus in {
            SolicitudCompra.STATUS_APROBADA,
            SolicitudCompra.STATUS_RECHAZADA,
        }
    )
    solicitud_ids = [item.id for item in solicitudes]
    orden_ready_ids = set()
    recepcion_ready_ids = set()
    if solicitud_ids:
        orden_ready_ids = set(
            OrdenCompra.objects.filter(solicitud_id__in=solicitud_ids)
            .values_list("solicitud_id", flat=True)
            .distinct()
        )
        recepcion_ready_ids = set(
            RecepcionCompra.objects.filter(
                orden__solicitud_id__in=solicitud_ids,
                estatus=RecepcionCompra.STATUS_CERRADA,
            )
            .values_list("orden__solicitud_id", flat=True)
            .distinct()
        )
    orden_ready = sum(1 for item in solicitudes if item.id in orden_ready_ids)
    recepcion_ready = sum(1 for item in solicitudes if item.id in recepcion_ready_ids)

    def _completion(closed: int) -> int:
        if total <= 0:
            return 0
        return int(round((closed / total) * 100))

    return [
        {
            "step": "01",
            "title": "Maestro listo",
            "closed": maestro_ready,
            "pending": max(total - maestro_ready, 0),
            "completion": _completion(maestro_ready),
            "detail": "Solicitud sin bloqueos ERP del artículo, costo y proveedor.",
            "url": "?estatus=BLOCKED_ERP",
            "cta": "Resolver bloqueos",
        },
        {
            "step": "02",
            "title": "Validación resuelta",
            "closed": validacion_ready,
            "pending": max(total - validacion_ready, 0),
            "completion": _completion(validacion_ready),
            "detail": "Solicitud ya revisada y fuera de captura o revisión.",
            "url": "?workflow_action=aprobar_rechazar",
            "cta": "Abrir validación",
        },
        {
            "step": "03",
            "title": "Orden emitida",
            "closed": orden_ready,
            "pending": max(total - orden_ready, 0),
            "completion": _completion(orden_ready),
            "detail": "Solicitud ya convertida en orden de compra activa.",
            "url": "?estatus=APPROVED_WITH_OC",
            "cta": "Abrir con OC",
        },
        {
            "step": "04",
            "title": "Recepción cerrada",
            "closed": recepcion_ready,
            "pending": max(total - recepcion_ready, 0),
            "completion": _completion(recepcion_ready),
            "detail": "Solicitud con ciclo documental cerrado hasta recepción.",
            "url": "?estatus=CERRADA",
            "cta": "Abrir cerradas",
        },
    ]


def _enrich_orden_workflow(orden, closed_recepciones=0):
    orden.workflow_blocker_details = _enterprise_blocker_details_for_orden(orden)
    orden.workflow_blockers = [item["message"] for item in orden.workflow_blocker_details]
    orden.has_workflow_blockers = bool(orden.workflow_blockers)
    orden.enterprise_master_summary = _enterprise_master_summary_from_details(
        _enterprise_blocker_details_for_solicitud(orden.solicitud) if orden.solicitud_id else []
    )
    if orden.estatus == OrdenCompra.STATUS_BORRADOR:
        orden.workflow_stage = "Preparación"
        orden.workflow_next = "Completar datos ERP" if orden.has_workflow_blockers else "Enviar a proveedor"
    elif orden.estatus == OrdenCompra.STATUS_ENVIADA:
        orden.workflow_stage = "Esperando confirmación"
        orden.workflow_next = "Confirmar o marcar parcial"
    elif orden.estatus == OrdenCompra.STATUS_CONFIRMADA:
        orden.workflow_stage = "Abastecimiento"
        orden.workflow_next = "Registrar recepción"
    elif orden.estatus == OrdenCompra.STATUS_PARCIAL:
        orden.workflow_stage = "Recepción parcial"
        orden.workflow_next = "Completar recepciones y cerrar"
    else:
        orden.workflow_stage = "Cerrada"
        orden.workflow_next = "Sin acción"
    orden.closed_recepciones = closed_recepciones
    if orden.has_workflow_blockers:
        orden.workflow_health_label = "Bloqueada ERP"
        orden.workflow_health_tone = "danger"
    elif orden.estatus == OrdenCompra.STATUS_ENVIADA:
        orden.workflow_health_label = "Esperando proveedor"
        orden.workflow_health_tone = "primary"
    elif orden.estatus == OrdenCompra.STATUS_CONFIRMADA:
        orden.workflow_health_label = "Lista para recepción"
        orden.workflow_health_tone = "success"
    elif orden.estatus == OrdenCompra.STATUS_PARCIAL:
        orden.workflow_health_label = "Recepción parcial"
        orden.workflow_health_tone = "warning"
    elif orden.estatus == OrdenCompra.STATUS_CERRADA:
        orden.workflow_health_label = "Cerrada"
        orden.workflow_health_tone = "muted"
    else:
        orden.workflow_health_label = "En preparación"
        orden.workflow_health_tone = "warning"

    if orden.has_workflow_blockers:
        orden.workflow_action_label = "Corregir datos ERP"
        orden.workflow_action_code = "corregir_datos"
    elif orden.estatus == OrdenCompra.STATUS_BORRADOR:
        orden.workflow_action_label = "Enviar a proveedor"
        orden.workflow_action_code = "enviar_proveedor"
    elif orden.estatus == OrdenCompra.STATUS_ENVIADA:
        orden.workflow_action_label = "Confirmar o parcial"
        orden.workflow_action_code = "confirmar_parcial"
    elif orden.estatus == OrdenCompra.STATUS_CONFIRMADA:
        orden.workflow_action_label = "Registrar recepción"
        orden.workflow_action_code = "registrar_recepcion"
    elif orden.estatus == OrdenCompra.STATUS_PARCIAL:
        orden.workflow_action_label = "Completar recepción"
        orden.workflow_action_code = "completar_recepcion"
    else:
        orden.workflow_action_label = "Sin acción"
        orden.workflow_action_code = "none"
    return orden


def _enterprise_blocker_details_for_orden(orden) -> list[dict[str, object]]:
    detail_rows: list[dict[str, object]] = []
    blocked_query = f"{reverse('compras:ordenes')}?estatus=BLOCKED_ERP&q={orden.folio}"
    if not orden.proveedor_id:
        detail_rows.append(
            {
                "key": "sin_proveedor",
                "message": "Sin proveedor",
                "missing_field": "proveedor",
                "action_label": "Completar proveedor",
                "action_url": blocked_query,
                "tone": "danger",
            }
        )
    if not orden.fecha_emision:
        detail_rows.append(
            {
                "key": "sin_emision",
                "message": "Sin fecha de emisión",
                "missing_field": "fecha emisión",
                "action_label": "Registrar emisión",
                "action_url": blocked_query,
                "tone": "danger",
            }
        )
    if (orden.monto_estimado or Decimal("0")) <= 0:
        detail_rows.append(
            {
                "key": "monto_cero",
                "message": "Monto estimado en cero",
                "missing_field": "monto",
                "action_label": "Corregir monto",
                "action_url": blocked_query,
                "tone": "warning",
            }
        )
    if orden.estatus == OrdenCompra.STATUS_BORRADOR and not orden.fecha_entrega_estimada:
        detail_rows.append(
            {
                "key": "sin_entrega",
                "message": "Sin fecha de entrega estimada",
                "missing_field": "fecha entrega",
                "action_label": "Registrar entrega",
                "action_url": blocked_query,
                "tone": "warning",
            }
        )
    return detail_rows


def _enterprise_blockers_for_orden(orden) -> list[str]:
    return [item["message"] for item in _enterprise_blocker_details_for_orden(orden)]


def _ordenes_workflow_summary(ordenes):
    blocked_erp = sum(1 for item in ordenes if getattr(item, "has_workflow_blockers", False))
    ready_for_recepcion = sum(
        1
        for item in ordenes
        if item.estatus in {OrdenCompra.STATUS_CONFIRMADA, OrdenCompra.STATUS_PARCIAL} and not getattr(item, "has_workflow_blockers", False)
    )
    return {
        "cards": _build_status_cards(
            ordenes,
            [
                (OrdenCompra.STATUS_BORRADOR, "Borrador", "warning"),
                (OrdenCompra.STATUS_ENVIADA, "Enviadas", "primary"),
                (OrdenCompra.STATUS_CONFIRMADA, "Confirmadas", "success"),
                (OrdenCompra.STATUS_PARCIAL, "Parciales", "warning"),
                (OrdenCompra.STATUS_CERRADA, "Cerradas", "success"),
            ],
        ),
        "gate_cards": [
            {
                "key": "erp_ready",
                "label": "ERP completo",
                "ready_count": max(len(ordenes) - blocked_erp, 0),
                "blocked_count": blocked_erp,
                "tone": "success" if blocked_erp == 0 else "danger",
                "detail": "Órdenes sin bloqueo de proveedor, emisión, monto o entrega estimada.",
                "action_label": "Corregir ERP",
                "query": "?workflow_action=corregir_datos",
            },
            {
                "key": "sent_or_confirmed",
                "label": "Proveedor atendido",
                "ready_count": sum(
                    1
                    for item in ordenes
                    if item.estatus in {
                        OrdenCompra.STATUS_ENVIADA,
                        OrdenCompra.STATUS_CONFIRMADA,
                        OrdenCompra.STATUS_PARCIAL,
                        OrdenCompra.STATUS_CERRADA,
                    }
                ),
                "blocked_count": sum(1 for item in ordenes if item.estatus == OrdenCompra.STATUS_BORRADOR),
                "tone": "success" if not any(item.estatus == OrdenCompra.STATUS_BORRADOR for item in ordenes) else "warning",
                "detail": "La orden ya salió de borrador y está en contacto con proveedor o en ejecución.",
                "action_label": "Abrir envío",
                "query": "?workflow_action=enviar_proveedor",
            },
            {
                "key": "ready_for_recepcion",
                "label": "Lista para recepción",
                "ready_count": ready_for_recepcion,
                "blocked_count": sum(1 for item in ordenes if item.estatus == OrdenCompra.STATUS_ENVIADA),
                "tone": "success" if ready_for_recepcion else "warning",
                "detail": "Órdenes confirmadas o parciales que pueden continuar a recepción.",
                "action_label": "Abrir recepción",
                "query": "?workflow_action=registrar_recepcion",
            },
        ],
        "note": "Flujo documental: Borrador -> Enviada -> Confirmada/Parcial -> Cerrada.",
    }


def _ordenes_enterprise_board(ordenes):
    next_step_defs = [
        ("corregir_datos", "Corregir datos ERP", "Corrección ERP", "danger"),
        ("enviar_proveedor", "Enviar a proveedor", "Enviar proveedor", "warning"),
        ("confirmar_parcial", "Confirmar o parcial", "Confirmación", "primary"),
        ("registrar_recepcion", "Registrar recepción", "Registrar recepción", "success"),
        ("completar_recepcion", "Completar recepción", "Completar recepción", "warning"),
    ]
    next_step_cards = []
    for key, action_label, label, tone in next_step_defs:
        next_step_cards.append(
            {
                "key": key,
                "label": label,
                "count": sum(1 for item in ordenes if getattr(item, "workflow_action_label", "") == action_label),
                "tone": tone,
                "query": f"?workflow_action={key}",
                "action_detail": f"Abrir órdenes en etapa {label.lower()}.",
            }
        )

    blocker_defs = [
        ("sin_proveedor", "Sin proveedor", "danger"),
        ("sin_emision", "Sin emisión", "danger"),
        ("monto_cero", "Monto en cero", "warning"),
        ("sin_entrega", "Sin entrega estimada", "warning"),
    ]
    blocker_counts = {key: 0 for key, _, _ in blocker_defs}
    for item in ordenes:
        for blocker in getattr(item, "workflow_blockers", []) or []:
            blocker_text = blocker.lower()
            if "sin proveedor" in blocker_text:
                blocker_counts["sin_proveedor"] += 1
            elif "sin fecha de emisión" in blocker_text or "sin fecha de emision" in blocker_text:
                blocker_counts["sin_emision"] += 1
            elif "monto estimado en cero" in blocker_text:
                blocker_counts["monto_cero"] += 1
            elif "sin fecha de entrega estimada" in blocker_text:
                blocker_counts["sin_entrega"] += 1
    blocker_cards = [
        {
            "key": key,
            "label": label,
            "count": blocker_counts[key],
            "tone": tone,
            "query": f"?blocker_key={key}",
            "action_detail": f"Abrir órdenes con bloqueo {label.lower()}.",
        }
        for key, label, tone in blocker_defs
    ]
    blocker_detail_rows = []
    for orden in ordenes:
        if not getattr(orden, "has_workflow_blockers", False):
            continue
        for blocker in getattr(orden, "workflow_blockers", []) or []:
            blocker_detail_rows.append(
                {
                    "folio": orden.folio,
                    "message": blocker,
                    "action_label": "Abrir orden",
                    "action_url": f"{reverse('compras:ordenes')}?estatus=BLOCKED_ERP&q={orden.folio}",
                }
            )
    master_rollup = _enterprise_master_blocker_rollup(
        [
            {
                "folio": orden.folio,
                "details": _enterprise_blocker_details_for_solicitud(orden.solicitud) if orden.solicitud_id else [],
            }
            for orden in ordenes
        ],
        usage_scope="purchases",
    )
    source_cards = [
        {
            "key": "manual",
            "label": "Manual",
            "count": sum(1 for item in ordenes if getattr(item, "source_tipo", "manual") == "manual"),
            "tone": "success",
            "query": "?source=manual",
        },
        {
            "key": "plan",
            "label": "Plan producción",
            "count": sum(1 for item in ordenes if getattr(item, "source_tipo", "") == "plan"),
            "tone": "warning",
            "query": "?source=plan",
        },
        {
            "key": "reabasto_cedis",
            "label": "Reabasto CEDIS",
            "count": sum(1 for item in ordenes if getattr(item, "source_tipo", "") == "reabasto_cedis"),
            "tone": "primary",
            "query": "?source=reabasto_cedis",
        },
    ]
    return {
        "next_step_cards": next_step_cards,
        "blocker_cards": blocker_cards,
        "blocker_detail_rows": blocker_detail_rows[:10],
        "master_blocker_class_cards": master_rollup["master_blocker_class_cards"],
        "master_blocker_missing_cards": master_rollup["master_blocker_missing_cards"],
        "master_blocker_detail_rows": master_rollup["master_blocker_detail_rows"],
        "source_cards": source_cards,
        "blocked_total": sum(1 for item in ordenes if getattr(item, "has_workflow_blockers", False)),
        "ready_for_recepcion_total": sum(1 for item in ordenes if item.estatus == OrdenCompra.STATUS_CONFIRMADA and not getattr(item, "has_workflow_blockers", False)),
    }


def _enrich_recepcion_workflow(recepcion):
    recepcion.workflow_blocker_details = _enterprise_blocker_details_for_recepcion(recepcion)
    recepcion.workflow_blockers = [item["message"] for item in recepcion.workflow_blocker_details]
    recepcion.has_workflow_blockers = bool(recepcion.workflow_blockers)
    recepcion.enterprise_master_summary = _enterprise_master_summary_from_details(
        _enterprise_blocker_details_for_solicitud(recepcion.orden.solicitud)
        if recepcion.orden_id and getattr(recepcion.orden, "solicitud_id", None)
        else []
    )
    if recepcion.estatus == RecepcionCompra.STATUS_PENDIENTE:
        recepcion.workflow_stage = "Validación"
        recepcion.workflow_next = "Corregir bloqueos ERP" if recepcion.has_workflow_blockers else "Cerrar o marcar diferencias"
    elif recepcion.estatus == RecepcionCompra.STATUS_DIFERENCIAS:
        recepcion.workflow_stage = "Resolver diferencias"
        recepcion.workflow_next = "Corregir bloqueos ERP" if recepcion.has_workflow_blockers else "Cerrar y aplicar inventario"
    else:
        recepcion.workflow_stage = "Aplicada a inventario"
        recepcion.workflow_next = "Sin acción"
    if recepcion.has_workflow_blockers:
        recepcion.workflow_health_label = "Bloqueada ERP"
        recepcion.workflow_health_tone = "danger"
    elif recepcion.estatus == RecepcionCompra.STATUS_DIFERENCIAS:
        recepcion.workflow_health_label = "Resolver diferencias"
        recepcion.workflow_health_tone = "warning"
    elif recepcion.estatus == RecepcionCompra.STATUS_CERRADA:
        recepcion.workflow_health_label = "Aplicada"
        recepcion.workflow_health_tone = "success"
    else:
        recepcion.workflow_health_label = "En validación"
        recepcion.workflow_health_tone = "primary"

    if recepcion.has_workflow_blockers:
        recepcion.workflow_action_label = "Corregir datos recepción"
        recepcion.workflow_action_code = "corregir_recepcion"
    elif recepcion.estatus == RecepcionCompra.STATUS_PENDIENTE:
        recepcion.workflow_action_label = "Cerrar o diferencias"
        recepcion.workflow_action_code = "cerrar_diferencias"
    elif recepcion.estatus == RecepcionCompra.STATUS_DIFERENCIAS:
        recepcion.workflow_action_label = "Cerrar y aplicar"
        recepcion.workflow_action_code = "cerrar_aplicar"
    else:
        recepcion.workflow_action_label = "Sin acción"
        recepcion.workflow_action_code = "none"
    return recepcion


def _recepciones_workflow_summary(recepciones):
    cards = _build_status_cards(
        recepciones,
        [
            (RecepcionCompra.STATUS_PENDIENTE, "Por validar", "warning"),
            (RecepcionCompra.STATUS_DIFERENCIAS, "Con diferencias", "danger"),
            (RecepcionCompra.STATUS_CERRADA, "Cerradas", "success"),
        ],
    )
    cards.append(
        {
            "code": "BLOCKED_ERP",
            "label": "Bloqueadas ERP",
            "count": sum(1 for r in recepciones if getattr(r, "has_workflow_blockers", False)),
            "tone": "danger",
            "query": "?estatus=BLOCKED_ERP",
        }
    )
    blocked_erp = sum(1 for r in recepciones if getattr(r, "has_workflow_blockers", False))
    applied_total = sum(1 for r in recepciones if r.estatus == RecepcionCompra.STATUS_CERRADA and not getattr(r, "has_workflow_blockers", False))
    return {
        "cards": cards,
        "gate_cards": [
            {
                "key": "erp_ready",
                "label": "Recepción válida",
                "ready_count": max(len(recepciones) - blocked_erp, 0),
                "blocked_count": blocked_erp,
                "tone": "success" if blocked_erp == 0 else "danger",
                "detail": "Recepciones con fecha, conformidad y justificación válidas para seguir a cierre.",
                "action_label": "Corregir recepción",
                "query": "?workflow_action=corregir_recepcion",
            },
            {
                "key": "differences_resolved",
                "label": "Diferencias resueltas",
                "ready_count": sum(1 for r in recepciones if r.estatus == RecepcionCompra.STATUS_CERRADA),
                "blocked_count": sum(1 for r in recepciones if r.estatus == RecepcionCompra.STATUS_DIFERENCIAS),
                "tone": "success" if not any(r.estatus == RecepcionCompra.STATUS_DIFERENCIAS for r in recepciones) else "warning",
                "detail": "Las recepciones ya no tienen diferencias abiertas antes del cierre.",
                "action_label": "Abrir diferencias",
                "query": "?estatus=diferencias",
            },
            {
                "key": "applied_inventory",
                "label": "Aplicada a inventario",
                "ready_count": applied_total,
                "blocked_count": sum(1 for r in recepciones if r.estatus != RecepcionCompra.STATUS_CERRADA),
                "tone": "success" if applied_total else "warning",
                "detail": "Recepciones cerradas y listas para reflejarse como evidencia de inventario.",
                "action_label": "Abrir aplicadas",
                "query": "?workflow_action=cerrar_aplicar",
            },
        ],
        "note": "Flujo documental: Por validar -> Diferencias -> Cerrada.",
    }


def _recepciones_enterprise_board(recepciones):
    next_step_defs = [
        ("corregir_recepcion", "Corregir datos recepción", "Corrección ERP", "danger"),
        ("cerrar_diferencias", "Cerrar o diferencias", "Validar recepción", "primary"),
        ("cerrar_aplicar", "Cerrar y aplicar", "Aplicar inventario", "warning"),
    ]
    next_step_cards = []
    for key, action_label, label, tone in next_step_defs:
        next_step_cards.append(
            {
                "key": key,
                "label": label,
                "count": sum(1 for item in recepciones if getattr(item, "workflow_action_label", "") == action_label),
                "tone": tone,
                "query": f"?workflow_action={key}",
                "action_detail": f"Abrir recepciones en etapa {label.lower()}.",
            }
        )

    blocker_defs = [
        ("sin_fecha", "Sin fecha", "danger"),
        ("conformidad", "Conformidad inválida", "danger"),
        ("sin_observacion", "Sin observación", "warning"),
        ("sin_justificacion", "Sin justificación", "warning"),
    ]
    blocker_counts = {key: 0 for key, _, _ in blocker_defs}
    for item in recepciones:
        for blocker in getattr(item, "workflow_blockers", []) or []:
            blocker_text = blocker.lower()
            if "sin fecha de recepción" in blocker_text or "sin fecha de recepcion" in blocker_text:
                blocker_counts["sin_fecha"] += 1
            elif "conformidad fuera de rango" in blocker_text:
                blocker_counts["conformidad"] += 1
            elif "sin observaciones de diferencia" in blocker_text:
                blocker_counts["sin_observacion"] += 1
            elif "sin justificación" in blocker_text or "sin justificacion" in blocker_text:
                blocker_counts["sin_justificacion"] += 1
    blocker_cards = [
        {
            "key": key,
            "label": label,
            "count": blocker_counts[key],
            "tone": tone,
            "query": f"?blocker_key={key}",
            "action_detail": f"Abrir recepciones con bloqueo {label.lower()}.",
        }
        for key, label, tone in blocker_defs
    ]
    blocker_detail_rows = []
    for recepcion in recepciones:
        if not getattr(recepcion, "has_workflow_blockers", False):
            continue
        for blocker in getattr(recepcion, "workflow_blockers", []) or []:
            blocker_detail_rows.append(
                {
                    "folio": recepcion.folio,
                    "message": blocker,
                    "action_label": "Abrir recepción",
                    "action_url": f"{reverse('compras:recepciones')}?estatus=BLOCKED_ERP&q={recepcion.folio}",
                }
            )
    master_rollup = _enterprise_master_blocker_rollup(
        [
            {
                "folio": recepcion.folio,
                "details": _enterprise_blocker_details_for_solicitud(recepcion.orden.solicitud)
                if recepcion.orden_id and getattr(recepcion.orden, "solicitud_id", None)
                else [],
            }
            for recepcion in recepciones
        ],
        usage_scope="purchases",
    )
    source_cards = [
        {
            "key": "manual",
            "label": "Manual",
            "count": sum(1 for item in recepciones if getattr(item, "source_tipo", "manual") == "manual"),
            "tone": "success",
            "query": "?source=manual",
        },
        {
            "key": "plan",
            "label": "Plan producción",
            "count": sum(1 for item in recepciones if getattr(item, "source_tipo", "") == "plan"),
            "tone": "warning",
            "query": "?source=plan",
        },
        {
            "key": "reabasto_cedis",
            "label": "Reabasto CEDIS",
            "count": sum(1 for item in recepciones if getattr(item, "source_tipo", "") == "reabasto_cedis"),
            "tone": "primary",
            "query": "?source=reabasto_cedis",
        },
    ]
    return {
        "next_step_cards": next_step_cards,
        "blocker_cards": blocker_cards,
        "blocker_detail_rows": blocker_detail_rows[:10],
        "master_blocker_class_cards": master_rollup["master_blocker_class_cards"],
        "master_blocker_missing_cards": master_rollup["master_blocker_missing_cards"],
        "master_blocker_detail_rows": master_rollup["master_blocker_detail_rows"],
        "source_cards": source_cards,
        "blocked_total": sum(1 for item in recepciones if getattr(item, "has_workflow_blockers", False)),
        "applied_total": sum(1 for item in recepciones if item.estatus == RecepcionCompra.STATUS_CERRADA and not getattr(item, "has_workflow_blockers", False)),
    }


def _enterprise_blockers_for_recepcion(recepcion) -> list[str]:
    return [item["message"] for item in _enterprise_blocker_details_for_recepcion(recepcion)]


def _enterprise_blocker_details_for_recepcion(recepcion) -> list[dict[str, object]]:
    blockers = []
    blocked_query = f"{reverse('compras:recepciones')}?estatus=BLOCKED_ERP&q={recepcion.folio}"
    if not recepcion.fecha_recepcion:
        blockers.append(
            {
                "key": "sin_fecha_recepcion",
                "message": "Sin fecha de recepción",
                "missing_field": "fecha recepción",
                "action_label": "Registrar fecha",
                "action_url": blocked_query,
                "tone": "danger",
            }
        )
    conformidad = _to_decimal(str(recepcion.conformidad_pct or 0), "0")
    if conformidad <= 0 or conformidad > 100:
        blockers.append(
            {
                "key": "conformidad_fuera_rango",
                "message": "Conformidad fuera de rango",
                "missing_field": "conformidad",
                "action_label": "Corregir conformidad",
                "action_url": blocked_query,
                "tone": "danger",
            }
        )
    observaciones = (recepcion.observaciones or "").strip()
    if recepcion.estatus == RecepcionCompra.STATUS_DIFERENCIAS and not observaciones:
        blockers.append(
            {
                "key": "sin_observaciones_diferencia",
                "message": "Sin observaciones de diferencia",
                "missing_field": "observaciones",
                "action_label": "Registrar observaciones",
                "action_url": blocked_query,
                "tone": "danger",
            }
        )
    if recepcion.estatus == RecepcionCompra.STATUS_CERRADA and conformidad < 100 and not observaciones:
        blockers.append(
            {
                "key": "sin_justificacion_conformidad",
                "message": "Conformidad menor a 100% sin justificación",
                "missing_field": "justificación",
                "action_label": "Registrar justificación",
                "action_url": blocked_query,
                "tone": "warning",
            }
        )
    return blockers


def _apply_recepcion_to_inventario(recepcion: RecepcionCompra, acted_by=None) -> dict:
    orden = recepcion.orden
    solicitud = orden.solicitud
    if not solicitud or not solicitud.insumo_id:
        return {"applied": False, "reason": "sin_solicitud_o_insumo"}

    cantidad = _to_decimal(str(solicitud.cantidad or 0), "0")
    if cantidad <= 0:
        return {"applied": False, "reason": "cantidad_no_positiva"}

    source_hash = f"recepcion:{recepcion.id}:entrada"
    if MovimientoInventario.objects.filter(source_hash=source_hash).exists():
        return {"applied": False, "reason": "ya_aplicado"}

    insumo_canonical = canonical_insumo_by_id(solicitud.insumo_id) or solicitud.insumo
    existencia, _ = ExistenciaInsumo.objects.get_or_create(insumo=insumo_canonical)
    prev_stock = existencia.stock_actual
    existencia.stock_actual = prev_stock + cantidad
    existencia.actualizado_en = timezone.now()
    existencia.save()

    movimiento = MovimientoInventario.objects.create(
        tipo=MovimientoInventario.TIPO_ENTRADA,
        insumo=insumo_canonical,
        cantidad=cantidad,
        referencia=recepcion.folio,
        source_hash=source_hash,
    )

    log_event(
        acted_by,
        "CREATE",
        "inventario.MovimientoInventario",
        movimiento.id,
        {
            "tipo": movimiento.tipo,
            "insumo_id": movimiento.insumo_id,
            "cantidad": str(movimiento.cantidad),
            "referencia": movimiento.referencia,
            "source": recepcion.folio,
        },
    )
    log_event(
        acted_by,
        "UPDATE",
        "inventario.ExistenciaInsumo",
        existencia.id,
        {
            "insumo_id": insumo_canonical.id,
            "from_stock": str(prev_stock),
            "to_stock": str(existencia.stock_actual),
            "source": recepcion.folio,
        },
    )
    return {"applied": True, "movimiento_id": movimiento.id}


def _build_insumo_options(limit: int = 1200):
    canonical_rows = canonicalized_active_insumos(limit=limit)
    grouped_member_ids = [member_id for row in canonical_rows for member_id in row["member_ids"]]
    usage_maps = usage_maps_for_insumo_ids(grouped_member_ids)
    existencias = {
        e.insumo_id: e
        for e in ExistenciaInsumo.objects.filter(insumo_id__in=grouped_member_ids).select_related("insumo", "insumo__proveedor_principal")
    }

    en_transito_by_insumo: dict[int, Decimal] = {}
    for orden in (
        OrdenCompra.objects.filter(
            estatus__in=[
                OrdenCompra.STATUS_ENVIADA,
                OrdenCompra.STATUS_CONFIRMADA,
                OrdenCompra.STATUS_PARCIAL,
            ],
            solicitud__isnull=False,
            solicitud__insumo_id__in=grouped_member_ids,
        )
        .select_related("solicitud")
        .only("solicitud__insumo_id", "solicitud__cantidad")
    ):
        solicitud = orden.solicitud
        if not solicitud or not solicitud.insumo_id:
            continue
        qty = _to_decimal(str(solicitud.cantidad or 0), "0")
        if qty <= 0:
            continue
        en_transito_by_insumo[solicitud.insumo_id] = en_transito_by_insumo.get(solicitud.insumo_id, Decimal("0")) + qty

    options = []
    for row in canonical_rows:
        insumo = row["canonical"]
        member_ids = row["member_ids"]
        member_existencias = [existencias[mid] for mid in member_ids if mid in existencias]
        ex = existencias.get(insumo.id) or next((item for item in member_existencias if item), None)
        stock_actual = sum((item.stock_actual for item in member_existencias), Decimal("0"))
        punto_reorden = ex.punto_reorden if ex else Decimal("0")
        stock_seguridad = ex.stock_minimo if ex else Decimal("0")
        consumo_diario = ex.consumo_diario_promedio if ex else Decimal("0")
        lead_time_dias = int(ex.dias_llegada_pedido or 0) if ex else 0
        if lead_time_dias <= 0 and insumo.proveedor_principal_id:
            lead_time_dias = int(insumo.proveedor_principal.lead_time_dias or 0)
        lead_time_dias = max(lead_time_dias, 0)

        demanda_lead_time = consumo_diario * Decimal(str(lead_time_dias))
        en_transito = sum((en_transito_by_insumo.get(member_id, Decimal("0")) for member_id in member_ids), Decimal("0"))
        recomendado = (demanda_lead_time + stock_seguridad) - (stock_actual + en_transito)
        if recomendado < 0:
            recomendado = Decimal("0")
        enterprise_profile = getattr(insumo, "enterprise_profile", enterprise_readiness_profile(insumo))
        purchase_count = sum(int(usage_maps["purchase_counts"].get(member_id, 0)) for member_id in member_ids)
        is_operational_blocker = enterprise_profile["readiness_label"] == "Incompleto" and purchase_count > 0

        options.append(
            {
                "id": insumo.id,
                "nombre": insumo.nombre,
                "canonical_variant_count": row["variant_count"],
                "proveedor_sugerido": insumo.proveedor_principal.nombre if insumo.proveedor_principal_id else "",
                "stock_actual": stock_actual,
                "punto_reorden": punto_reorden,
                "stock_seguridad": stock_seguridad,
                "demanda_lead_time": demanda_lead_time,
                "en_transito": en_transito,
                "lead_time_dias": lead_time_dias,
                "recomendado": recomendado,
                "enterprise_status": enterprise_profile["readiness_label"],
                "enterprise_missing": enterprise_profile["missing"],
                "is_operational_blocker": is_operational_blocker,
                "operational_blocker_label": "Bloquea compras" if is_operational_blocker else "",
            }
        )
    return options


def _canonical_catalog_maps(limit: int = 5000) -> tuple[dict[int, dict], dict[int, dict]]:
    canonical_rows = canonicalized_active_insumos(limit=limit)
    member_to_row: dict[int, dict] = {}
    canonical_by_id: dict[int, dict] = {}
    for row in canonical_rows:
        canonical = row["canonical"]
        canonical_by_id[canonical.id] = row
        for member_id in row["member_ids"]:
            member_to_row[member_id] = row
    return member_to_row, canonical_by_id


def _latest_cost_by_canonical_ids(canonical_ids: set[int], canonical_by_id: dict[int, dict]) -> dict[int, Decimal]:
    latest_cost_by_canonical: dict[int, Decimal] = {}
    for canonical_id in canonical_ids:
        latest = latest_costo_canonico(insumo_id=canonical_id)
        if latest is not None:
            latest_cost_by_canonical[canonical_id] = latest
    return latest_cost_by_canonical


def _solicitudes_print_folio() -> str:
    now = timezone.localtime()
    return f"SC-{now.strftime('%Y%m%d-%H%M%S')}"


def _parse_period_filters(periodo_tipo_raw: str, periodo_mes_raw: str) -> tuple[str, str, str]:
    tipo = (periodo_tipo_raw or "all").strip().lower()
    if tipo not in {"all", "mes", "q1", "q2"}:
        tipo = "all"

    now = timezone.localdate()
    default_mes = f"{now.year:04d}-{now.month:02d}"
    periodo_mes = (periodo_mes_raw or default_mes).strip()
    try:
        y, m = periodo_mes.split("-")
        y_int = int(y)
        m_int = int(m)
        if not (1 <= m_int <= 12):
            raise ValueError
        periodo_mes = f"{y_int:04d}-{m_int:02d}"
    except Exception:
        periodo_mes = default_mes

    if tipo == "mes":
        label = f"Mensual ({periodo_mes})"
    elif tipo == "q1":
        label = f"1ra Quincena ({periodo_mes})"
    elif tipo == "q2":
        label = f"2da Quincena ({periodo_mes})"
    else:
        label = "Todos"
    return tipo, periodo_mes, label


def _periodo_bounds(periodo_tipo: str, periodo_mes: str) -> tuple[date | None, date | None]:
    if periodo_tipo == "all":
        return None, None

    year, month = periodo_mes.split("-")
    y = int(year)
    m = int(month)
    last_day = calendar.monthrange(y, m)[1]
    start = date(y, m, 1)
    end = date(y, m, last_day)

    if periodo_tipo == "q1":
        end = date(y, m, 15)
    elif periodo_tipo == "q2":
        start = date(y, m, 16)
    return start, end


def _filter_ordenes_by_scope(ordenes_qs, source_filter: str, plan_filter: str):
    if source_filter == "plan":
        ordenes_qs = ordenes_qs.filter(
            Q(solicitud__area__startswith="PLAN_PRODUCCION:")
            | Q(referencia__startswith="PLAN_PRODUCCION:")
        )
    elif source_filter == "manual":
        ordenes_qs = ordenes_qs.exclude(
            Q(solicitud__area__startswith="PLAN_PRODUCCION:")
            | Q(referencia__startswith="PLAN_PRODUCCION:")
        )

    if plan_filter:
        plan_scope = f"PLAN_PRODUCCION:{plan_filter}"
        ordenes_qs = ordenes_qs.filter(
            Q(solicitud__area=plan_scope)
            | Q(referencia=plan_scope)
        )
    return ordenes_qs


def _filter_recepciones_by_scope(recepciones_qs, source_filter: str, plan_filter: str):
    if source_filter == "plan":
        recepciones_qs = recepciones_qs.filter(
            Q(orden__solicitud__area__startswith="PLAN_PRODUCCION:")
            | Q(orden__referencia__startswith="PLAN_PRODUCCION:")
        )
    elif source_filter == "manual":
        recepciones_qs = recepciones_qs.exclude(
            Q(orden__solicitud__area__startswith="PLAN_PRODUCCION:")
            | Q(orden__referencia__startswith="PLAN_PRODUCCION:")
        )

    if plan_filter:
        plan_scope = f"PLAN_PRODUCCION:{plan_filter}"
        recepciones_qs = recepciones_qs.filter(
            Q(orden__solicitud__area=plan_scope)
            | Q(orden__referencia=plan_scope)
        )
    return recepciones_qs


def _filter_documents_by_master_blockers(documents, article_class_filter: str, missing_field_filter: str):
    article_class_filter = (article_class_filter or "all").strip()
    missing_field_filter = (missing_field_filter or "all").strip()
    if article_class_filter == "all" and missing_field_filter == "all":
        return list(documents)

    filtered = []
    for document in documents:
        details = getattr(document, "enterprise_master_blocker_details", []) or []
        if not details:
            continue
        matched = False
        for detail in details:
            insumo = detail.get("insumo")
            if insumo is None and detail.get("insumo_id"):
                insumo = canonical_insumo_by_id(detail.get("insumo_id"))
            class_key = _enterprise_article_class(insumo).get("key")
            missing_key = _missing_field_filter_key(detail.get("missing_field")) or "other"
            class_ok = article_class_filter == "all" or class_key == article_class_filter
            missing_ok = missing_field_filter == "all" or missing_key == missing_field_filter
            if class_ok and missing_ok:
                matched = True
                break
        if matched:
            filtered.append(document)
    return filtered


def _solicitud_has_blocker_key(solicitud, blocker_key: str) -> bool:
    text_values = " ".join(getattr(solicitud, "workflow_blockers", []) or []).lower()
    if blocker_key == "sin_costo":
        return "sin costo vigente" in text_values
    if blocker_key == "sin_proveedor":
        return "sin proveedor sugerido" in text_values
    if blocker_key == "maestro_incompleto":
        return "artículo incompleto" in text_values or "articulo incompleto" in text_values
    if blocker_key == "articulo_inactivo":
        return "artículo inactivo" in text_values or "articulo inactivo" in text_values
    if blocker_key == "sin_catalogo":
        return "catálogo canónico" in text_values or "catalogo canonico" in text_values
    if blocker_key == "no_canonico":
        return "variante no canónica" in text_values or "variante no canonica" in text_values
    return False


def _orden_has_blocker_key(orden, blocker_key: str) -> bool:
    text_values = " ".join(getattr(orden, "workflow_blockers", []) or []).lower()
    if blocker_key == "sin_proveedor":
        return "sin proveedor" in text_values
    if blocker_key == "sin_emision":
        return "sin fecha de emisión" in text_values or "sin fecha de emision" in text_values
    if blocker_key == "monto_cero":
        return "monto estimado en cero" in text_values
    if blocker_key == "sin_entrega":
        return "sin fecha de entrega estimada" in text_values
    return False


def _recepcion_has_blocker_key(recepcion, blocker_key: str) -> bool:
    text_values = " ".join(getattr(recepcion, "workflow_blockers", []) or []).lower()
    if blocker_key == "sin_fecha":
        return "sin fecha de recepción" in text_values or "sin fecha de recepcion" in text_values
    if blocker_key == "conformidad":
        return "conformidad fuera de rango" in text_values
    if blocker_key == "sin_observacion":
        return "sin observaciones de diferencia" in text_values
    if blocker_key == "sin_justificacion":
        return "sin justificación" in text_values or "sin justificacion" in text_values
    return False


def _filter_solicitudes_by_scope(solicitudes_qs, source_filter: str, plan_filter: str):
    if source_filter == "plan":
        solicitudes_qs = solicitudes_qs.filter(area__startswith="PLAN_PRODUCCION:")
    elif source_filter == "manual":
        solicitudes_qs = solicitudes_qs.exclude(area__startswith="PLAN_PRODUCCION:")
    if plan_filter:
        solicitudes_qs = solicitudes_qs.filter(area=f"PLAN_PRODUCCION:{plan_filter}")
    return solicitudes_qs


def _filter_solicitudes_by_categoria(solicitudes_qs, categoria_filter: str):
    categoria = _sanitize_categoria_filter(categoria_filter)
    if not categoria:
        return solicitudes_qs

    categoria_norm = _normalize_categoria_text(categoria)
    categoria_unit_map = {
        "masa": "MASS",
        "volumen": "VOLUME",
        "pieza": "UNIT",
    }
    unidad_tipo = categoria_unit_map.get(categoria_norm)
    if unidad_tipo:
        return solicitudes_qs.filter(
            Q(insumo__categoria__iexact=categoria)
            | (
                (Q(insumo__categoria="") | Q(insumo__categoria__isnull=True))
                & Q(insumo__unidad_base__tipo=unidad_tipo)
            )
        )
    return solicitudes_qs.filter(insumo__categoria__iexact=categoria)


def _filter_ordenes_by_categoria(ordenes_qs, categoria_filter: str):
    categoria = _sanitize_categoria_filter(categoria_filter)
    if not categoria:
        return ordenes_qs

    categoria_norm = _normalize_categoria_text(categoria)
    categoria_unit_map = {
        "masa": "MASS",
        "volumen": "VOLUME",
        "pieza": "UNIT",
    }
    unidad_tipo = categoria_unit_map.get(categoria_norm)
    if unidad_tipo:
        return ordenes_qs.filter(
            Q(solicitud__insumo__categoria__iexact=categoria)
            | (
                (Q(solicitud__insumo__categoria="") | Q(solicitud__insumo__categoria__isnull=True))
                & Q(solicitud__insumo__unidad_base__tipo=unidad_tipo)
            )
        )
    return ordenes_qs.filter(solicitud__insumo__categoria__iexact=categoria)


def _filter_movimientos_by_scope(
    movimientos_qs,
    source_filter: str,
    plan_filter: str,
    consumo_ref_filter: str = "all",
):
    if source_filter == "plan":
        movimientos_qs = movimientos_qs.filter(referencia__icontains="PLAN_PRODUCCION:")
    elif source_filter == "manual":
        movimientos_qs = movimientos_qs.exclude(referencia__icontains="PLAN_PRODUCCION:")
    if plan_filter:
        movimientos_qs = movimientos_qs.filter(referencia__icontains=f"PLAN_PRODUCCION:{plan_filter}")
    if _sanitize_consumo_ref_filter(consumo_ref_filter) == "plan_ref":
        movimientos_qs = movimientos_qs.filter(referencia__icontains="PLAN_PRODUCCION:")
    return movimientos_qs


def _filter_movimientos_by_categoria(movimientos_qs, categoria_filter: str):
    categoria = _sanitize_categoria_filter(categoria_filter)
    if not categoria:
        return movimientos_qs

    categoria_norm = _normalize_categoria_text(categoria)
    categoria_unit_map = {
        "masa": "MASS",
        "volumen": "VOLUME",
        "pieza": "UNIT",
    }
    unidad_tipo = categoria_unit_map.get(categoria_norm)
    if unidad_tipo:
        return movimientos_qs.filter(
            Q(insumo__categoria__iexact=categoria)
            | (
                (Q(insumo__categoria="") | Q(insumo__categoria__isnull=True))
                & Q(insumo__unidad_base__tipo=unidad_tipo)
            )
        )
    return movimientos_qs.filter(insumo__categoria__iexact=categoria)


def _shift_month(periodo_mes: str, delta_months: int) -> str:
    year, month = periodo_mes.split("-")
    y = int(year)
    m = int(month)
    total = y * 12 + (m - 1) - delta_months
    shifted_y = total // 12
    shifted_m = (total % 12) + 1
    return f"{shifted_y:04d}-{shifted_m:02d}"


def _compute_budget_period_summary(
    periodo_tipo: str,
    periodo_mes: str,
    source_filter: str,
    plan_filter: str,
    categoria_filter: str,
) -> dict:
    start_date, end_date = _periodo_bounds(periodo_tipo, periodo_mes)
    if not start_date or not end_date:
        return {
            "periodo_tipo": periodo_tipo,
            "periodo_mes": periodo_mes,
            "objetivo": Decimal("0"),
            "estimado": Decimal("0"),
            "ejecutado": Decimal("0"),
            "ratio_pct": None,
            "estado_label": "Sin periodo",
            "estado_badge": "bg-warning",
        }

    presupuesto = PresupuestoCompraPeriodo.objects.filter(
        periodo_tipo=periodo_tipo,
        periodo_mes=periodo_mes,
    ).first()
    objetivo = presupuesto.monto_objetivo if presupuesto else Decimal("0")

    solicitudes_qs = SolicitudCompra.objects.filter(fecha_requerida__range=(start_date, end_date))
    solicitudes_qs = _filter_solicitudes_by_scope(solicitudes_qs, source_filter, plan_filter)
    solicitudes_qs = _filter_solicitudes_by_categoria(solicitudes_qs, categoria_filter)
    solicitudes_vals = list(solicitudes_qs.values("insumo_id", "cantidad"))
    total_qty_by_canonical: dict[int, Decimal] = {}
    _, canonical_by_id = _canonical_catalog_maps()
    for row in solicitudes_vals:
        canonical = canonical_insumo_by_id(row["insumo_id"])
        if not canonical:
            continue
        total_qty_by_canonical[canonical.id] = total_qty_by_canonical.get(canonical.id, Decimal("0")) + (
            row.get("cantidad") or Decimal("0")
        )

    latest_cost_by_insumo = _latest_cost_by_canonical_ids(set(total_qty_by_canonical.keys()), canonical_by_id)
    estimado = sum(
        qty * latest_cost_by_insumo.get(canonical_id, Decimal("0"))
        for canonical_id, qty in total_qty_by_canonical.items()
    )

    ordenes_qs = OrdenCompra.objects.exclude(estatus=OrdenCompra.STATUS_BORRADOR).filter(
        fecha_emision__range=(start_date, end_date)
    )
    ordenes_qs = _filter_ordenes_by_scope(ordenes_qs, source_filter, plan_filter)
    ordenes_qs = _filter_ordenes_by_categoria(ordenes_qs, categoria_filter)
    ejecutado = ordenes_qs.aggregate(total=Sum("monto_estimado"))["total"] or Decimal("0")

    base = max(estimado, ejecutado)
    ratio_pct = ((base * Decimal("100")) / objetivo) if objetivo > 0 else None
    if objetivo <= 0:
        estado_label = "Sin objetivo"
        estado_badge = "bg-warning"
    elif ratio_pct <= Decimal("90"):
        estado_label = "Verde"
        estado_badge = "bg-success"
    elif ratio_pct <= Decimal("100"):
        estado_label = "Amarillo"
        estado_badge = "bg-warning"
    else:
        estado_label = "Rojo"
        estado_badge = "bg-danger"

    return {
        "periodo_tipo": periodo_tipo,
        "periodo_mes": periodo_mes,
        "objetivo": objetivo,
        "estimado": estimado,
        "ejecutado": ejecutado,
        "ratio_pct": ratio_pct,
        "estado_label": estado_label,
        "estado_badge": estado_badge,
    }


def _build_budget_history(periodo_mes: str, source_filter: str, plan_filter: str, categoria_filter: str) -> list[dict]:
    rows: list[dict] = []
    for delta in range(0, 6):
        month_value = _shift_month(periodo_mes, delta)
        summary = _compute_budget_period_summary("mes", month_value, source_filter, plan_filter, categoria_filter)
        rows.append(summary)
    return rows


def _build_provider_dashboard(
    periodo_mes: str,
    source_filter: str,
    plan_filter: str,
    categoria_filter: str,
    current_rows: list[dict],
) -> dict:
    months_desc = [_shift_month(periodo_mes, d) for d in range(0, 6)]
    months_asc = list(reversed(months_desc))

    monthly_provider_data: dict[str, dict[str, dict[str, Decimal]]] = {}
    provider_score: dict[str, Decimal] = {}

    for month_value in months_desc:
        start_date, end_date = _periodo_bounds("mes", month_value)
        solicitudes_qs = SolicitudCompra.objects.select_related("insumo", "proveedor_sugerido").filter(
            fecha_requerida__range=(start_date, end_date)
        )
        solicitudes_qs = _filter_solicitudes_by_scope(solicitudes_qs, source_filter, plan_filter)
        solicitudes_qs = _filter_solicitudes_by_categoria(solicitudes_qs, categoria_filter)
        solicitudes = list(solicitudes_qs)

        _, canonical_by_id = _canonical_catalog_maps()
        canonical_ids = {canonical_insumo_by_id(s.insumo_id).id for s in solicitudes if canonical_insumo_by_id(s.insumo_id)}
        latest_cost_by_insumo = _latest_cost_by_canonical_ids(canonical_ids, canonical_by_id)

        estimado_by_provider: dict[str, Decimal] = {}
        for s in solicitudes:
            proveedor_nombre = (
                s.proveedor_sugerido.nombre
                if s.proveedor_sugerido_id
                else (
                    s.insumo.proveedor_principal.nombre
                    if getattr(s.insumo, "proveedor_principal_id", None)
                    else "Sin proveedor"
                )
            )
            canonical = canonical_insumo_by_id(s.insumo_id)
            estimado_by_provider[proveedor_nombre] = estimado_by_provider.get(proveedor_nombre, Decimal("0")) + (
                (s.cantidad or Decimal("0")) * latest_cost_by_insumo.get(canonical.id, Decimal("0"))
            )

        ordenes_qs = OrdenCompra.objects.exclude(estatus=OrdenCompra.STATUS_BORRADOR).filter(
            fecha_emision__range=(start_date, end_date)
        )
        ordenes_qs = _filter_ordenes_by_scope(ordenes_qs, source_filter, plan_filter)
        ordenes_qs = _filter_ordenes_by_categoria(ordenes_qs, categoria_filter)
        ejecutado_by_provider: dict[str, Decimal] = {}
        for row in ordenes_qs.values("proveedor__nombre").annotate(total=Sum("monto_estimado")):
            provider_name = row["proveedor__nombre"] or "Sin proveedor"
            ejecutado_by_provider[provider_name] = row["total"] or Decimal("0")

        providers = set(estimado_by_provider.keys()) | set(ejecutado_by_provider.keys())
        for provider_name in providers:
            estimado = estimado_by_provider.get(provider_name, Decimal("0"))
            ejecutado = ejecutado_by_provider.get(provider_name, Decimal("0"))
            variacion = ejecutado - estimado
            monthly_provider_data.setdefault(provider_name, {})[month_value] = {
                "estimado": estimado,
                "ejecutado": ejecutado,
                "variacion": variacion,
            }
            provider_score[provider_name] = provider_score.get(provider_name, Decimal("0")) + abs(variacion)

    top_providers = [p for p, _ in sorted(provider_score.items(), key=lambda x: x[1], reverse=True)[:6]]

    trend_rows: list[dict] = []
    for provider_name in top_providers:
        for month_value in months_asc:
            data = monthly_provider_data.get(provider_name, {}).get(
                month_value,
                {"estimado": Decimal("0"), "ejecutado": Decimal("0"), "variacion": Decimal("0")},
            )
            trend_rows.append(
                {
                    "proveedor": provider_name,
                    "mes": month_value,
                    "estimado": data["estimado"],
                    "ejecutado": data["ejecutado"],
                    "variacion": data["variacion"],
                }
            )

    top_desviaciones = sorted(
        [r for r in current_rows if (r["variacion"] or Decimal("0")) != Decimal("0")],
        key=lambda x: abs(x["variacion"]),
        reverse=True,
    )[:12]

    return {
        "top_desviaciones": top_desviaciones,
        "trend_rows": trend_rows,
        "trend_months": months_asc,
        "trend_providers": top_providers,
    }


def _build_category_dashboard(
    periodo_mes: str,
    source_filter: str,
    plan_filter: str,
    categoria_filter: str,
    current_rows: list[dict],
) -> dict:
    months_desc = [_shift_month(periodo_mes, d) for d in range(0, 6)]
    months_asc = list(reversed(months_desc))

    monthly_category_data: dict[str, dict[str, dict[str, Decimal]]] = {}
    category_score: dict[str, Decimal] = {}

    for month_value in months_desc:
        start_date, end_date = _periodo_bounds("mes", month_value)
        solicitudes_qs = (
            SolicitudCompra.objects.select_related("insumo", "insumo__unidad_base", "proveedor_sugerido")
            .filter(fecha_requerida__range=(start_date, end_date))
        )
        solicitudes_qs = _filter_solicitudes_by_scope(solicitudes_qs, source_filter, plan_filter)
        solicitudes_qs = _filter_solicitudes_by_categoria(solicitudes_qs, categoria_filter)
        solicitudes = list(solicitudes_qs)

        _, canonical_by_id = _canonical_catalog_maps()
        canonical_ids = {canonical_insumo_by_id(s.insumo_id).id for s in solicitudes if canonical_insumo_by_id(s.insumo_id)}
        latest_cost_by_insumo = _latest_cost_by_canonical_ids(canonical_ids, canonical_by_id)

        estimado_by_categoria: dict[str, Decimal] = {}
        for s in solicitudes:
            canonical = canonical_insumo_by_id(s.insumo_id) or s.insumo
            categoria_nombre = _resolve_insumo_categoria(canonical)
            estimado = getattr(s, "presupuesto_estimado", None)
            if estimado is None:
                estimado = (s.cantidad or Decimal("0")) * latest_cost_by_insumo.get(canonical.id, Decimal("0"))
            estimado_by_categoria[categoria_nombre] = estimado_by_categoria.get(categoria_nombre, Decimal("0")) + (
                estimado or Decimal("0")
            )

        ordenes_qs = (
            OrdenCompra.objects.select_related("solicitud", "solicitud__insumo", "solicitud__insumo__unidad_base")
            .exclude(estatus=OrdenCompra.STATUS_BORRADOR)
            .filter(fecha_emision__range=(start_date, end_date))
        )
        ordenes_qs = _filter_ordenes_by_scope(ordenes_qs, source_filter, plan_filter)
        ordenes_qs = _filter_ordenes_by_categoria(ordenes_qs, categoria_filter)
        ejecutado_by_categoria: dict[str, Decimal] = {}
        for orden in ordenes_qs:
            categoria_nombre = "Sin categoría"
            if orden.solicitud_id and getattr(orden.solicitud, "insumo_id", None):
                categoria_nombre = _resolve_insumo_categoria(
                    canonical_insumo_by_id(orden.solicitud.insumo_id) or orden.solicitud.insumo
                )
            ejecutado_by_categoria[categoria_nombre] = ejecutado_by_categoria.get(categoria_nombre, Decimal("0")) + (
                orden.monto_estimado or Decimal("0")
            )

        categorias = set(estimado_by_categoria.keys()) | set(ejecutado_by_categoria.keys())
        for categoria_nombre in categorias:
            estimado = estimado_by_categoria.get(categoria_nombre, Decimal("0"))
            ejecutado = ejecutado_by_categoria.get(categoria_nombre, Decimal("0"))
            variacion = ejecutado - estimado
            monthly_category_data.setdefault(categoria_nombre, {})[month_value] = {
                "estimado": estimado,
                "ejecutado": ejecutado,
                "variacion": variacion,
            }
            category_score[categoria_nombre] = category_score.get(categoria_nombre, Decimal("0")) + abs(variacion)

    top_categorias = [c for c, _ in sorted(category_score.items(), key=lambda x: x[1], reverse=True)[:6]]

    trend_rows: list[dict] = []
    for categoria_nombre in top_categorias:
        for month_value in months_asc:
            data = monthly_category_data.get(categoria_nombre, {}).get(
                month_value,
                {"estimado": Decimal("0"), "ejecutado": Decimal("0"), "variacion": Decimal("0")},
            )
            trend_rows.append(
                {
                    "categoria": categoria_nombre,
                    "mes": month_value,
                    "estimado": data["estimado"],
                    "ejecutado": data["ejecutado"],
                    "variacion": data["variacion"],
                }
            )

    top_desviaciones = sorted(
        [r for r in current_rows if (r["variacion"] or Decimal("0")) != Decimal("0")],
        key=lambda x: abs(x["variacion"]),
        reverse=True,
    )[:12]

    return {
        "top_desviaciones": top_desviaciones,
        "trend_rows": trend_rows,
        "trend_months": months_asc,
        "trend_categories": top_categorias,
    }


def _build_consumo_vs_plan_dashboard(
    periodo_tipo: str,
    periodo_mes: str,
    source_filter: str,
    plan_filter: str,
    categoria_filter: str,
    consumo_ref_filter: str = "all",
    *,
    limit: int = 30,
    offset: int = 0,
    sort_by: str = "variacion_cost_abs",
    sort_dir: str = "desc",
) -> dict:
    start_date, end_date = _periodo_bounds(periodo_tipo, periodo_mes)
    consumo_ref_filter = _sanitize_consumo_ref_filter(consumo_ref_filter)
    if not start_date or not end_date:
        # Evita consultas históricas abiertas cuando el filtro está en "Todos".
        end_date = timezone.localdate()
        start_date = end_date - timedelta(days=90)

    plan_qs = PlanProduccion.objects.prefetch_related(
        "items__receta__lineas__insumo__unidad_base"
    ).filter(fecha_produccion__range=(start_date, end_date))
    if plan_filter and plan_filter.isdigit():
        plan_qs = plan_qs.filter(id=int(plan_filter))
    if source_filter == "manual":
        plan_qs = plan_qs.none()

    plan_qty_by_insumo: dict[int, Decimal] = {}
    insumo_meta: dict[int, dict] = {}
    for plan in plan_qs:
        for item in plan.items.all():
            multiplicador = Decimal(str(item.cantidad or 0))
            if multiplicador <= 0:
                continue
            for linea in item.receta.lineas.all():
                if not linea.insumo_id:
                    continue
                qty_base = Decimal(str(linea.cantidad or 0))
                if qty_base <= 0:
                    continue
                qty_requerida = qty_base * multiplicador
                if qty_requerida <= 0:
                    continue
                if categoria_filter and _sanitize_categoria_filter(categoria_filter):
                    if _normalize_categoria_text(_resolve_insumo_categoria(linea.insumo)) != _normalize_categoria_text(
                        _sanitize_categoria_filter(categoria_filter)
                    ):
                        continue
                plan_qty_by_insumo[linea.insumo_id] = plan_qty_by_insumo.get(linea.insumo_id, Decimal("0")) + qty_requerida
                if linea.insumo_id not in insumo_meta:
                    insumo_meta[linea.insumo_id] = {
                        "insumo": linea.insumo.nombre,
                        "unidad": linea.insumo.unidad_base.codigo if linea.insumo.unidad_base_id else "-",
                        "categoria": _resolve_insumo_categoria(linea.insumo),
                    }

    movimientos_qs = (
        MovimientoInventario.objects.select_related("insumo", "insumo__unidad_base")
        .filter(tipo__in=[MovimientoInventario.TIPO_SALIDA, MovimientoInventario.TIPO_CONSUMO])
        .filter(fecha__date__range=(start_date, end_date))
    )
    movimientos_qs = _filter_movimientos_by_scope(
        movimientos_qs,
        source_filter,
        plan_filter,
        consumo_ref_filter,
    )
    movimientos_qs = _filter_movimientos_by_categoria(movimientos_qs, categoria_filter)

    actual_qty_by_insumo: dict[int, Decimal] = {}
    for mov in movimientos_qs:
        if not mov.insumo_id:
            continue
        qty = abs(Decimal(str(mov.cantidad or 0)))
        if qty <= 0:
            continue
        actual_qty_by_insumo[mov.insumo_id] = actual_qty_by_insumo.get(mov.insumo_id, Decimal("0")) + qty
        if mov.insumo_id not in insumo_meta:
            insumo_meta[mov.insumo_id] = {
                "insumo": mov.insumo.nombre,
                "unidad": mov.insumo.unidad_base.codigo if mov.insumo.unidad_base_id else "-",
                "categoria": _resolve_insumo_categoria(mov.insumo),
            }

    member_to_row, canonical_by_id = _canonical_catalog_maps()
    plan_qty_by_canonical: dict[int, Decimal] = {}
    actual_qty_by_canonical: dict[int, Decimal] = {}
    canonical_meta: dict[int, dict] = {}
    for insumo_id, qty in plan_qty_by_insumo.items():
        row = member_to_row.get(insumo_id)
        if not row:
            continue
        canonical = row["canonical"]
        plan_qty_by_canonical[canonical.id] = plan_qty_by_canonical.get(canonical.id, Decimal("0")) + qty
        canonical_meta.setdefault(canonical.id, insumo_meta.get(insumo_id, {
            "insumo": canonical.nombre,
            "unidad": canonical.unidad_base.codigo if getattr(canonical, "unidad_base_id", None) else "-",
            "categoria": _resolve_insumo_categoria(canonical),
        }))
    for insumo_id, qty in actual_qty_by_insumo.items():
        row = member_to_row.get(insumo_id)
        if not row:
            continue
        canonical = row["canonical"]
        actual_qty_by_canonical[canonical.id] = actual_qty_by_canonical.get(canonical.id, Decimal("0")) + qty
        canonical_meta.setdefault(canonical.id, insumo_meta.get(insumo_id, {
            "insumo": canonical.nombre,
            "unidad": canonical.unidad_base.codigo if getattr(canonical, "unidad_base_id", None) else "-",
            "categoria": _resolve_insumo_categoria(canonical),
        }))

    insumo_ids = list(set(plan_qty_by_canonical.keys()) | set(actual_qty_by_canonical.keys()))
    latest_cost_by_insumo = _latest_cost_by_canonical_ids(set(insumo_ids), canonical_by_id)

    rows: list[dict] = []
    totals = {
        "plan_qty_total": Decimal("0"),
        "consumo_real_qty_total": Decimal("0"),
        "plan_cost_total": Decimal("0"),
        "consumo_real_cost_total": Decimal("0"),
        "variacion_cost_total": Decimal("0"),
        "sin_costo_count": 0,
        "semaforo_verde_count": 0,
        "semaforo_amarillo_count": 0,
        "semaforo_rojo_count": 0,
    }
    for insumo_id in insumo_ids:
        plan_qty = plan_qty_by_canonical.get(insumo_id, Decimal("0"))
        real_qty = actual_qty_by_canonical.get(insumo_id, Decimal("0"))
        costo_unitario = latest_cost_by_insumo.get(insumo_id, Decimal("0"))

        plan_cost = plan_qty * costo_unitario
        real_cost = real_qty * costo_unitario
        variacion_qty = real_qty - plan_qty
        variacion_cost = real_cost - plan_cost
        consumo_pct = None
        if plan_qty > 0:
            consumo_pct = (real_qty * Decimal("100")) / plan_qty
        sin_costo = costo_unitario <= 0 and (plan_qty > 0 or real_qty > 0)

        estado = "OK"
        semaforo = "VERDE"
        if plan_qty <= 0 and real_qty > 0:
            estado = "SIN_PLAN"
            semaforo = "ROJO"
        elif plan_qty > 0 and real_qty <= 0:
            estado = "SIN_CONSUMO"
            semaforo = "AMARILLO"
        elif consumo_pct is not None and consumo_pct > Decimal("110"):
            estado = "SOBRECONSUMO"
            semaforo = "ROJO"
        elif consumo_pct is not None and consumo_pct < Decimal("90"):
            estado = "BAJO_CONSUMO"
            semaforo = "AMARILLO"

        if sin_costo and semaforo == "VERDE":
            semaforo = "AMARILLO"

        if semaforo == "ROJO":
            totals["semaforo_rojo_count"] += 1
        elif semaforo == "AMARILLO":
            totals["semaforo_amarillo_count"] += 1
        else:
            totals["semaforo_verde_count"] += 1

        if sin_costo:
            totals["sin_costo_count"] += 1

        meta = canonical_meta.get(insumo_id, {"insumo": f"Insumo {insumo_id}", "unidad": "-", "categoria": "Sin categoría"})
        rows.append(
            {
                "insumo_id": insumo_id,
                "insumo": meta["insumo"],
                "categoria": meta["categoria"],
                "unidad": meta["unidad"],
                "cantidad_plan": plan_qty,
                "cantidad_real": real_qty,
                "variacion_qty": variacion_qty,
                "costo_unitario": costo_unitario,
                "costo_plan": plan_cost,
                "costo_real": real_cost,
                "variacion_cost": variacion_cost,
                "consumo_pct": consumo_pct,
                "estado": estado,
                "semaforo": semaforo,
                "sin_costo": sin_costo,
                "alerta": "Sin costo unitario" if sin_costo else "",
            }
        )
        totals["plan_qty_total"] += plan_qty
        totals["consumo_real_qty_total"] += real_qty
        totals["plan_cost_total"] += plan_cost
        totals["consumo_real_cost_total"] += real_cost
        totals["variacion_cost_total"] += variacion_cost

    try:
        limit = int(limit)
    except (TypeError, ValueError):
        limit = 30
    limit = max(1, min(limit, 1000))

    try:
        offset = int(offset)
    except (TypeError, ValueError):
        offset = 0
    offset = max(0, min(offset, 200000))

    allowed_sort = {
        "variacion_cost_abs",
        "variacion_cost",
        "costo_real",
        "costo_plan",
        "cantidad_real",
        "cantidad_plan",
        "consumo_pct",
        "insumo",
        "categoria",
        "estado",
        "semaforo",
    }
    sort_by = (sort_by or "variacion_cost_abs").strip().lower()
    if sort_by not in allowed_sort:
        sort_by = "variacion_cost_abs"
    sort_dir = (sort_dir or "desc").strip().lower()
    if sort_dir not in {"asc", "desc"}:
        sort_dir = "desc"

    def _safe_text(value: str | None) -> str:
        return (value or "").strip().lower()

    sort_map = {
        "variacion_cost_abs": lambda row: abs(row.get("variacion_cost") or Decimal("0")),
        "variacion_cost": lambda row: row.get("variacion_cost") or Decimal("0"),
        "costo_real": lambda row: row.get("costo_real") or Decimal("0"),
        "costo_plan": lambda row: row.get("costo_plan") or Decimal("0"),
        "cantidad_real": lambda row: row.get("cantidad_real") or Decimal("0"),
        "cantidad_plan": lambda row: row.get("cantidad_plan") or Decimal("0"),
        "consumo_pct": lambda row: row.get("consumo_pct") if row.get("consumo_pct") is not None else Decimal("-1"),
        "insumo": lambda row: _safe_text(row.get("insumo")),
        "categoria": lambda row: _safe_text(row.get("categoria")),
        "estado": lambda row: _safe_text(row.get("estado")),
        "semaforo": lambda row: _safe_text(row.get("semaforo")),
    }
    rows_sorted = sorted(rows, key=sort_map[sort_by], reverse=(sort_dir == "desc"))
    rows_total = len(rows_sorted)
    rows_paginated = rows_sorted[offset : offset + limit]

    totals["cobertura_pct"] = (
        (totals["consumo_real_qty_total"] * Decimal("100")) / totals["plan_qty_total"]
        if totals["plan_qty_total"] > 0
        else None
    )
    return {
        "rows": rows_paginated,
        "totals": totals,
        "period_start": start_date,
        "period_end": end_date,
        "meta": {
            "rows_total": rows_total,
            "rows_returned": len(rows_paginated),
            "limit": limit,
            "offset": offset,
            "sort_by": sort_by,
            "sort_dir": sort_dir,
        },
    }


def _build_budget_context(
    solicitudes,
    source_filter: str,
    plan_filter: str,
    categoria_filter: str,
    periodo_tipo: str,
    periodo_mes: str,
) -> dict:
    def _estado_objetivo_label(estado: str) -> str:
        if estado == "excedido":
            return "Excedido"
        if estado == "preventivo":
            return "Preventivo"
        if estado == "ok":
            return "OK"
        return "Sin objetivo"

    total_estimado = sum((s.presupuesto_estimado for s in solicitudes), Decimal("0"))

    start_date, end_date = _periodo_bounds(periodo_tipo, periodo_mes)
    ordenes_qs = (
        OrdenCompra.objects.select_related("proveedor", "solicitud", "solicitud__insumo", "solicitud__insumo__unidad_base")
        .exclude(estatus=OrdenCompra.STATUS_BORRADOR)
    )
    if start_date and end_date:
        ordenes_qs = ordenes_qs.filter(fecha_emision__range=(start_date, end_date))
    ordenes_qs = _filter_ordenes_by_scope(ordenes_qs, source_filter, plan_filter)
    ordenes_qs = _filter_ordenes_by_categoria(ordenes_qs, categoria_filter)

    total_ejecutado = ordenes_qs.aggregate(total=Sum("monto_estimado"))["total"] or Decimal("0")
    variacion_ejecutado_vs_estimado = total_ejecutado - total_estimado

    presupuesto_periodo = None
    objetivo = None
    variacion_objetivo = None
    variacion_objetivo_pct = None
    avance_objetivo_pct = None
    objetivos_proveedor_by_name: dict[str, PresupuestoCompraProveedor] = {}
    objetivos_categoria_by_norm: dict[str, PresupuestoCompraCategoria] = {}
    if periodo_tipo != "all":
        presupuesto_periodo = PresupuestoCompraPeriodo.objects.filter(
            periodo_tipo=periodo_tipo,
            periodo_mes=periodo_mes,
        ).first()
        objetivo = presupuesto_periodo.monto_objetivo if presupuesto_periodo else Decimal("0")
        variacion_objetivo = total_estimado - objetivo
        if objetivo > 0:
            variacion_objetivo_pct = (variacion_objetivo * Decimal("100")) / objetivo
            avance_objetivo_pct = (total_ejecutado * Decimal("100")) / objetivo
        if presupuesto_periodo:
            for objetivo_prov in (
                PresupuestoCompraProveedor.objects.select_related("proveedor")
                .filter(presupuesto_periodo=presupuesto_periodo)
                .only("id", "proveedor__nombre", "monto_objetivo")
            ):
                objetivos_proveedor_by_name[objetivo_prov.proveedor.nombre] = objetivo_prov
            for objetivo_cat in (
                PresupuestoCompraCategoria.objects.filter(presupuesto_periodo=presupuesto_periodo)
                .only("id", "categoria", "categoria_normalizada", "monto_objetivo")
            ):
                objetivos_categoria_by_norm[objetivo_cat.categoria_normalizada] = objetivo_cat

    estimado_by_proveedor: dict[str, Decimal] = {}
    estimado_by_categoria: dict[str, Decimal] = {}
    for s in solicitudes:
        canonical = canonical_insumo_by_id(s.insumo_id) or s.insumo
        proveedor_nombre = (
            s.proveedor_sugerido.nombre
            if s.proveedor_sugerido_id
            else (
                canonical.proveedor_principal.nombre
                if getattr(canonical, "proveedor_principal_id", None)
                else "Sin proveedor"
            )
        )
        estimado_by_proveedor[proveedor_nombre] = estimado_by_proveedor.get(proveedor_nombre, Decimal("0")) + (
            s.presupuesto_estimado or Decimal("0")
        )
        categoria_nombre = _resolve_insumo_categoria(canonical)
        estimado_by_categoria[categoria_nombre] = estimado_by_categoria.get(categoria_nombre, Decimal("0")) + (
            s.presupuesto_estimado or Decimal("0")
        )

    ejecutado_by_proveedor: dict[str, Decimal] = {}
    for row in ordenes_qs.values("proveedor__nombre").annotate(total=Sum("monto_estimado")):
        proveedor_nombre = row["proveedor__nombre"] or "Sin proveedor"
        ejecutado_by_proveedor[proveedor_nombre] = row["total"] or Decimal("0")

    ejecutado_by_categoria: dict[str, Decimal] = {}
    for orden in ordenes_qs:
        categoria_nombre = "Sin categoría"
        if orden.solicitud_id and getattr(orden.solicitud, "insumo_id", None):
            categoria_nombre = _resolve_insumo_categoria(
                canonical_insumo_by_id(orden.solicitud.insumo_id) or orden.solicitud.insumo
            )
        ejecutado_by_categoria[categoria_nombre] = ejecutado_by_categoria.get(categoria_nombre, Decimal("0")) + (
            orden.monto_estimado or Decimal("0")
        )

    proveedores = (
        set(estimado_by_proveedor.keys())
        | set(ejecutado_by_proveedor.keys())
        | set(objetivos_proveedor_by_name.keys())
    )
    rows_proveedor = []
    proveedor_estado_counts = {
        "excedido": 0,
        "preventivo": 0,
        "ok": 0,
        "sin_objetivo": 0,
    }
    for proveedor_nombre in proveedores:
        estimado = estimado_by_proveedor.get(proveedor_nombre, Decimal("0"))
        ejecutado = ejecutado_by_proveedor.get(proveedor_nombre, Decimal("0"))
        variacion = ejecutado - estimado
        objetivo_proveedor_obj = objetivos_proveedor_by_name.get(proveedor_nombre)
        objetivo_proveedor = (
            objetivo_proveedor_obj.monto_objetivo if objetivo_proveedor_obj else Decimal("0")
        )
        base_control = max(estimado, ejecutado)
        uso_objetivo_pct = (
            (base_control * Decimal("100")) / objetivo_proveedor
            if objetivo_proveedor > 0
            else None
        )
        objetivo_estado = "sin_objetivo"
        if objetivo_proveedor > 0:
            if base_control > objetivo_proveedor:
                objetivo_estado = "excedido"
            elif base_control >= (objetivo_proveedor * Decimal("0.90")):
                objetivo_estado = "preventivo"
            else:
                objetivo_estado = "ok"
        proveedor_estado_counts[objetivo_estado] = proveedor_estado_counts.get(objetivo_estado, 0) + 1
        share = (estimado * Decimal("100") / total_estimado) if total_estimado > 0 else Decimal("0")
        rows_proveedor.append(
            {
                "proveedor": proveedor_nombre,
                "estimado": estimado,
                "ejecutado": ejecutado,
                "variacion": variacion,
                "participacion_pct": share,
                "objetivo_proveedor": objetivo_proveedor,
                "uso_objetivo_pct": uso_objetivo_pct,
                "objetivo_estado": objetivo_estado,
                "objetivo_estado_label": _estado_objetivo_label(objetivo_estado),
            }
        )
    rows_proveedor.sort(
        key=lambda r: max(
            r["estimado"] or Decimal("0"),
            r["ejecutado"] or Decimal("0"),
            r["objetivo_proveedor"] or Decimal("0"),
        ),
        reverse=True,
    )

    categorias = (
        set(estimado_by_categoria.keys())
        | set(ejecutado_by_categoria.keys())
        | set(
            obj.categoria
            for obj in objetivos_categoria_by_norm.values()
            if (obj.categoria or "").strip()
        )
    )
    rows_categoria = []
    categoria_estado_counts = {
        "excedido": 0,
        "preventivo": 0,
        "ok": 0,
        "sin_objetivo": 0,
    }
    for categoria_nombre in categorias:
        categoria_display = " ".join((categoria_nombre or "").strip().split()) or "Sin categoría"
        categoria_norm = _normalize_categoria_text(categoria_display)
        estimado = estimado_by_categoria.get(categoria_display, Decimal("0"))
        ejecutado = ejecutado_by_categoria.get(categoria_display, Decimal("0"))
        variacion = ejecutado - estimado
        objetivo_categoria_obj = objetivos_categoria_by_norm.get(categoria_norm)
        objetivo_categoria = (
            objetivo_categoria_obj.monto_objetivo if objetivo_categoria_obj else Decimal("0")
        )
        base_control = max(estimado, ejecutado)
        uso_objetivo_pct = (
            (base_control * Decimal("100")) / objetivo_categoria
            if objetivo_categoria > 0
            else None
        )
        objetivo_estado = "sin_objetivo"
        if objetivo_categoria > 0:
            if base_control > objetivo_categoria:
                objetivo_estado = "excedido"
            elif base_control >= (objetivo_categoria * Decimal("0.90")):
                objetivo_estado = "preventivo"
            else:
                objetivo_estado = "ok"
        categoria_estado_counts[objetivo_estado] = categoria_estado_counts.get(objetivo_estado, 0) + 1
        share = (estimado * Decimal("100") / total_estimado) if total_estimado > 0 else Decimal("0")
        rows_categoria.append(
            {
                "categoria": categoria_display,
                "estimado": estimado,
                "ejecutado": ejecutado,
                "variacion": variacion,
                "participacion_pct": share,
                "objetivo_categoria": objetivo_categoria,
                "uso_objetivo_pct": uso_objetivo_pct,
                "objetivo_estado": objetivo_estado,
                "objetivo_estado_label": _estado_objetivo_label(objetivo_estado),
            }
        )
    rows_categoria.sort(
        key=lambda r: max(
            r["estimado"] or Decimal("0"),
            r["ejecutado"] or Decimal("0"),
            r["objetivo_categoria"] or Decimal("0"),
        ),
        reverse=True,
    )

    alertas: list[dict] = []
    if periodo_tipo != "all":
        if objetivo is not None and objetivo > 0:
            if total_estimado > objetivo:
                alertas.append(
                    {
                        "nivel": "alto",
                        "tipo": "periodo_estimado",
                        "titulo": "Estimado supera objetivo",
                        "detalle": f"Estimado ${total_estimado:.2f} > Objetivo ${objetivo:.2f}",
                    }
                )
            if total_ejecutado > objetivo:
                alertas.append(
                    {
                        "nivel": "alto",
                        "tipo": "periodo_ejecutado",
                        "titulo": "Ejecutado supera objetivo",
                        "detalle": f"Ejecutado ${total_ejecutado:.2f} > Objetivo ${objetivo:.2f}",
                    }
                )

        for row in rows_proveedor:
            estimado = row["estimado"] or Decimal("0")
            ejecutado = row["ejecutado"] or Decimal("0")
            variacion = row["variacion"] or Decimal("0")
            objetivo_proveedor = row["objetivo_proveedor"] or Decimal("0")
            uso_objetivo_pct = row["uso_objetivo_pct"]
            if objetivo_proveedor > 0 and uso_objetivo_pct is not None:
                if uso_objetivo_pct > Decimal("100"):
                    alertas.append(
                        {
                            "nivel": "alto",
                            "tipo": "proveedor_objetivo_excedido",
                            "titulo": f"{row['proveedor']}: supera objetivo proveedor",
                            "detalle": f"${max(estimado, ejecutado):.2f} > ${objetivo_proveedor:.2f} ({uso_objetivo_pct:.2f}%)",
                        }
                    )
                elif uso_objetivo_pct >= Decimal("90"):
                    alertas.append(
                        {
                            "nivel": "medio",
                            "tipo": "proveedor_objetivo_preventivo",
                            "titulo": f"{row['proveedor']}: cerca de objetivo proveedor",
                            "detalle": f"${max(estimado, ejecutado):.2f} de ${objetivo_proveedor:.2f} ({uso_objetivo_pct:.2f}%)",
                        }
                    )
            if ejecutado <= 0:
                continue
            if estimado <= 0 and ejecutado > 0:
                alertas.append(
                    {
                        "nivel": "medio",
                        "tipo": "proveedor_sin_base",
                        "titulo": f"{row['proveedor']}: sin base estimada",
                        "detalle": f"Ejecutado ${ejecutado:.2f} sin estimado en solicitudes",
                    }
                )
                continue
            if variacion > 0:
                pct = (variacion * Decimal("100")) / estimado if estimado > 0 else Decimal("0")
                alertas.append(
                    {
                        "nivel": "medio",
                        "tipo": "proveedor_desviado",
                        "titulo": f"{row['proveedor']}: ejecutado arriba de estimado",
                        "detalle": f"+${variacion:.2f} ({pct:.2f}%) sobre estimado",
                    }
                )

        for row in rows_categoria:
            estimado = row["estimado"] or Decimal("0")
            ejecutado = row["ejecutado"] or Decimal("0")
            variacion = row["variacion"] or Decimal("0")
            objetivo_categoria = row["objetivo_categoria"] or Decimal("0")
            uso_objetivo_pct = row["uso_objetivo_pct"]
            if objetivo_categoria > 0 and uso_objetivo_pct is not None:
                if uso_objetivo_pct > Decimal("100"):
                    alertas.append(
                        {
                            "nivel": "alto",
                            "tipo": "categoria_objetivo_excedido",
                            "titulo": f"{row['categoria']}: supera objetivo categoría",
                            "detalle": f"${max(estimado, ejecutado):.2f} > ${objetivo_categoria:.2f} ({uso_objetivo_pct:.2f}%)",
                        }
                    )
                elif uso_objetivo_pct >= Decimal("90"):
                    alertas.append(
                        {
                            "nivel": "medio",
                            "tipo": "categoria_objetivo_preventivo",
                            "titulo": f"{row['categoria']}: cerca de objetivo categoría",
                            "detalle": f"${max(estimado, ejecutado):.2f} de ${objetivo_categoria:.2f} ({uso_objetivo_pct:.2f}%)",
                        }
                    )
            if ejecutado <= 0:
                continue
            if estimado <= 0 and ejecutado > 0:
                alertas.append(
                    {
                        "nivel": "medio",
                        "tipo": "categoria_sin_base",
                        "titulo": f"{row['categoria']}: sin base estimada",
                        "detalle": f"Ejecutado ${ejecutado:.2f} sin estimado en solicitudes",
                    }
                )
                continue
            if variacion > 0:
                pct = (variacion * Decimal("100")) / estimado if estimado > 0 else Decimal("0")
                alertas.append(
                    {
                        "nivel": "medio",
                        "tipo": "categoria_desviada",
                        "titulo": f"{row['categoria']}: ejecutado arriba de estimado",
                        "detalle": f"+${variacion:.2f} ({pct:.2f}%) sobre estimado",
                    }
                )
    alertas.sort(key=lambda x: (0 if x["nivel"] == "alto" else 1, x["titulo"]))

    return {
        "presupuesto_periodo": presupuesto_periodo,
        "presupuesto_objetivo": objetivo,
        "presupuesto_estimado_total": total_estimado,
        "presupuesto_ejecutado_total": total_ejecutado,
        "presupuesto_variacion_objetivo": variacion_objetivo,
        "presupuesto_variacion_objetivo_pct": variacion_objetivo_pct,
        "presupuesto_avance_objetivo_pct": avance_objetivo_pct,
        "presupuesto_variacion_ejecutado_estimado": variacion_ejecutado_vs_estimado,
        "presupuesto_rows_proveedor": rows_proveedor,
        "presupuesto_rows_categoria": rows_categoria,
        "presupuesto_objetivos_proveedor_total": len(objetivos_proveedor_by_name),
        "presupuesto_objetivos_categoria_total": len(objetivos_categoria_by_norm),
        "presupuesto_alertas": alertas[:25],
        "presupuesto_alertas_total": len(alertas),
        "presupuesto_alertas_altas": sum(1 for a in alertas if a["nivel"] == "alto"),
        "presupuesto_alertas_medias": sum(1 for a in alertas if a["nivel"] == "medio"),
        "presupuesto_alertas_preventivas": sum(
            1
            for a in alertas
            if a["tipo"] in {"proveedor_objetivo_preventivo", "categoria_objetivo_preventivo"}
        ),
        "presupuesto_alertas_excedidas": sum(
            1
            for a in alertas
            if a["tipo"]
            in {
                "proveedor_objetivo_excedido",
                "categoria_objetivo_excedido",
                "periodo_estimado",
                "periodo_ejecutado",
            }
        ),
        "presupuesto_proveedor_excedido_count": proveedor_estado_counts.get("excedido", 0),
        "presupuesto_proveedor_preventivo_count": proveedor_estado_counts.get("preventivo", 0),
        "presupuesto_proveedor_ok_count": proveedor_estado_counts.get("ok", 0),
        "presupuesto_proveedor_sin_objetivo_count": proveedor_estado_counts.get("sin_objetivo", 0),
        "presupuesto_categoria_excedido_count": categoria_estado_counts.get("excedido", 0),
        "presupuesto_categoria_preventivo_count": categoria_estado_counts.get("preventivo", 0),
        "presupuesto_categoria_ok_count": categoria_estado_counts.get("ok", 0),
        "presupuesto_categoria_sin_objetivo_count": categoria_estado_counts.get("sin_objetivo", 0),
    }


def _export_solicitudes_csv(
    solicitudes,
    source_filter: str,
    plan_filter: str,
    categoria_filter: str,
    reabasto_filter: str,
    periodo_tipo: str,
    periodo_mes: str,
    periodo_label: str,
) -> HttpResponse:
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="solicitudes_compras_{now_str}.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "Folio",
            "Area",
            "Solicitante",
            "Origen",
            "Plan",
            "Insumo",
            "Proveedor sugerido",
            "Cantidad",
            "Costo unitario",
            "Presupuesto estimado",
            "Fecha requerida",
            "Estatus",
            "Reabasto",
            "Detalle reabasto",
            "Filtro origen",
            "Filtro plan",
            "Filtro categoria",
            "Filtro reabasto",
            "Filtro periodo",
            "Filtro mes",
        ]
    )
    for s in solicitudes:
        if s.source_tipo == "reabasto_cedis":
            source_label = "REABASTO_CEDIS"
        elif s.source_tipo == "plan":
            source_label = "PLAN"
        else:
            source_label = "MANUAL"
        writer.writerow(
            [
                s.folio,
                s.area,
                s.solicitante,
                source_label,
                s.source_plan_id or "",
                s.insumo.nombre,
                s.proveedor_sugerido.nombre if s.proveedor_sugerido_id else "",
                s.cantidad,
                s.costo_unitario,
                s.presupuesto_estimado,
                s.fecha_requerida,
                s.get_estatus_display(),
                s.reabasto_texto,
                s.reabasto_detalle,
                source_filter,
                plan_filter or "",
                categoria_filter or "",
                reabasto_filter,
                periodo_label,
                periodo_mes if periodo_tipo != "all" else "",
            ]
        )
    return response


def _export_solicitudes_xlsx(
    solicitudes,
    source_filter: str,
    plan_filter: str,
    categoria_filter: str,
    reabasto_filter: str,
    periodo_tipo: str,
    periodo_mes: str,
    periodo_label: str,
) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes"
    ws.append(
        [
            "Folio",
            "Area",
            "Solicitante",
            "Origen",
            "Plan",
            "Insumo",
            "Proveedor sugerido",
            "Cantidad",
            "Costo unitario",
            "Presupuesto estimado",
            "Fecha requerida",
            "Estatus",
            "Reabasto",
            "Detalle reabasto",
            "Filtro origen",
            "Filtro plan",
            "Filtro categoria",
            "Filtro reabasto",
            "Filtro periodo",
            "Filtro mes",
        ]
    )
    for s in solicitudes:
        if s.source_tipo == "reabasto_cedis":
            source_label = "REABASTO_CEDIS"
        elif s.source_tipo == "plan":
            source_label = "PLAN"
        else:
            source_label = "MANUAL"
        ws.append(
            [
                s.folio,
                s.area,
                s.solicitante,
                source_label,
                s.source_plan_id or "",
                s.insumo.nombre,
                s.proveedor_sugerido.nombre if s.proveedor_sugerido_id else "",
                float(s.cantidad or 0),
                float(s.costo_unitario or 0),
                float(s.presupuesto_estimado or 0),
                s.fecha_requerida.isoformat() if s.fecha_requerida else "",
                s.get_estatus_display(),
                s.reabasto_texto,
                s.reabasto_detalle,
                source_filter,
                plan_filter or "",
                categoria_filter or "",
                reabasto_filter,
                periodo_label,
                periodo_mes if periodo_tipo != "all" else "",
            ]
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="solicitudes_compras_{now_str}.xlsx"'
    return response


def _export_consolidado_csv(
    solicitudes,
    source_filter: str,
    plan_filter: str,
    categoria_filter: str,
    reabasto_filter: str,
    periodo_label: str,
    budget_ctx: dict,
) -> HttpResponse:
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="compras_consolidado_{now_str}.csv"'
    writer = csv.writer(response)

    writer.writerow(["RESUMEN EJECUTIVO COMPRAS"])
    writer.writerow(["Filtro periodo", periodo_label])
    writer.writerow(["Filtro origen", source_filter])
    writer.writerow(["Filtro plan", plan_filter or "-"])
    writer.writerow(["Filtro categoria", categoria_filter or "-"])
    writer.writerow(["Filtro reabasto", reabasto_filter])
    writer.writerow(["Objetivo presupuesto", budget_ctx.get("presupuesto_objetivo") or ""])
    writer.writerow(["Estimado solicitudes", budget_ctx["presupuesto_estimado_total"]])
    writer.writerow(["Ejecutado ordenes", budget_ctx["presupuesto_ejecutado_total"]])
    writer.writerow(["Variacion vs objetivo", budget_ctx.get("presupuesto_variacion_objetivo") or ""])
    writer.writerow(["Variacion ejecutado vs estimado", budget_ctx["presupuesto_variacion_ejecutado_estimado"]])
    writer.writerow(["Proveedores excedidos", budget_ctx.get("presupuesto_proveedor_excedido_count") or 0])
    writer.writerow(["Proveedores preventivos", budget_ctx.get("presupuesto_proveedor_preventivo_count") or 0])
    writer.writerow(["Categorias excedidas", budget_ctx.get("presupuesto_categoria_excedido_count") or 0])
    writer.writerow(["Categorias preventivas", budget_ctx.get("presupuesto_categoria_preventivo_count") or 0])
    writer.writerow([])
    writer.writerow(["ALERTAS"])
    writer.writerow(["Nivel", "Tipo", "Titulo", "Detalle"])
    for alerta in budget_ctx.get("presupuesto_alertas", []):
        writer.writerow([alerta.get("nivel"), alerta.get("tipo"), alerta.get("titulo"), alerta.get("detalle")])
    writer.writerow([])

    writer.writerow(["DESVIACION POR PROVEEDOR"])
    writer.writerow(
        [
            "Proveedor",
            "Estimado",
            "Ejecutado",
            "Variacion",
            "Participacion estimado %",
            "Objetivo proveedor",
            "% Uso objetivo proveedor",
            "Estado objetivo",
        ]
    )
    for row in budget_ctx["presupuesto_rows_proveedor"]:
        writer.writerow(
            [
                row["proveedor"],
                row["estimado"],
                row["ejecutado"],
                row["variacion"],
                round(float(row["participacion_pct"]), 2),
                row.get("objetivo_proveedor", Decimal("0")),
                round(float(row["uso_objetivo_pct"] or 0), 2) if row.get("uso_objetivo_pct") is not None else "",
                row.get("objetivo_estado_label", ""),
            ]
        )
    writer.writerow([])

    writer.writerow(["DESVIACION POR CATEGORIA"])
    writer.writerow(
        [
            "Categoria",
            "Estimado",
            "Ejecutado",
            "Variacion",
            "Participacion estimado %",
            "Objetivo categoria",
            "% Uso objetivo categoria",
            "Estado objetivo",
        ]
    )
    for row in budget_ctx.get("presupuesto_rows_categoria", []):
        writer.writerow(
            [
                row["categoria"],
                row["estimado"],
                row["ejecutado"],
                row["variacion"],
                round(float(row["participacion_pct"]), 2),
                row.get("objetivo_categoria", Decimal("0")),
                round(float(row["uso_objetivo_pct"] or 0), 2) if row.get("uso_objetivo_pct") is not None else "",
                row.get("objetivo_estado_label", ""),
            ]
        )
    writer.writerow([])

    writer.writerow(["DETALLE SOLICITUDES"])
    writer.writerow(
        [
            "Folio",
            "Area",
            "Solicitante",
            "Insumo",
            "Proveedor sugerido",
            "Cantidad",
            "Costo unitario",
            "Presupuesto",
            "Fecha requerida",
            "Estatus",
        ]
    )
    for s in solicitudes:
        writer.writerow(
            [
                s.folio,
                s.area,
                s.solicitante,
                s.insumo.nombre,
                s.proveedor_sugerido.nombre if s.proveedor_sugerido_id else "",
                s.cantidad,
                s.costo_unitario,
                s.presupuesto_estimado,
                s.fecha_requerida,
                s.get_estatus_display(),
            ]
        )
    return response


def _export_consolidado_xlsx(
    solicitudes,
    source_filter: str,
    plan_filter: str,
    categoria_filter: str,
    reabasto_filter: str,
    periodo_label: str,
    budget_ctx: dict,
) -> HttpResponse:
    wb = Workbook()

    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen.append(["RESUMEN EJECUTIVO COMPRAS"])
    ws_resumen.append(["Filtro periodo", periodo_label])
    ws_resumen.append(["Filtro origen", source_filter])
    ws_resumen.append(["Filtro plan", plan_filter or "-"])
    ws_resumen.append(["Filtro categoria", categoria_filter or "-"])
    ws_resumen.append(["Filtro reabasto", reabasto_filter])
    ws_resumen.append(["Objetivo presupuesto", float(budget_ctx["presupuesto_objetivo"] or 0)])
    ws_resumen.append(["Estimado solicitudes", float(budget_ctx["presupuesto_estimado_total"] or 0)])
    ws_resumen.append(["Ejecutado ordenes", float(budget_ctx["presupuesto_ejecutado_total"] or 0)])
    ws_resumen.append(["Variacion vs objetivo", float((budget_ctx.get("presupuesto_variacion_objetivo") or 0))])
    ws_resumen.append(["Variacion ejecutado vs estimado", float(budget_ctx["presupuesto_variacion_ejecutado_estimado"] or 0)])
    ws_resumen.append(["Proveedores excedidos", int(budget_ctx.get("presupuesto_proveedor_excedido_count") or 0)])
    ws_resumen.append(["Proveedores preventivos", int(budget_ctx.get("presupuesto_proveedor_preventivo_count") or 0)])
    ws_resumen.append(["Categorias excedidas", int(budget_ctx.get("presupuesto_categoria_excedido_count") or 0)])
    ws_resumen.append(["Categorias preventivas", int(budget_ctx.get("presupuesto_categoria_preventivo_count") or 0)])
    ws_resumen.append([])
    ws_resumen.append(["ALERTAS"])
    ws_resumen.append(["Nivel", "Tipo", "Titulo", "Detalle"])
    for alerta in budget_ctx.get("presupuesto_alertas", []):
        ws_resumen.append(
            [
                alerta.get("nivel"),
                alerta.get("tipo"),
                alerta.get("titulo"),
                alerta.get("detalle"),
            ]
        )
    ws_resumen.append([])
    ws_resumen.append(["DESVIACION POR PROVEEDOR"])
    ws_resumen.append(
        [
            "Proveedor",
            "Estimado",
            "Ejecutado",
            "Variacion",
            "Participacion estimado %",
            "Objetivo proveedor",
            "% Uso objetivo proveedor",
            "Estado objetivo",
        ]
    )
    for row in budget_ctx["presupuesto_rows_proveedor"]:
        ws_resumen.append(
            [
                row["proveedor"],
                float(row["estimado"] or 0),
                float(row["ejecutado"] or 0),
                float(row["variacion"] or 0),
                float(row["participacion_pct"] or 0),
                float(row.get("objetivo_proveedor") or 0),
                float(row["uso_objetivo_pct"] or 0) if row.get("uso_objetivo_pct") is not None else None,
                row.get("objetivo_estado_label", ""),
            ]
        )
    ws_resumen.append([])
    ws_resumen.append(["DESVIACION POR CATEGORIA"])
    ws_resumen.append(
        [
            "Categoria",
            "Estimado",
            "Ejecutado",
            "Variacion",
            "Participacion estimado %",
            "Objetivo categoria",
            "% Uso objetivo categoria",
            "Estado objetivo",
        ]
    )
    for row in budget_ctx.get("presupuesto_rows_categoria", []):
        ws_resumen.append(
            [
                row["categoria"],
                float(row["estimado"] or 0),
                float(row["ejecutado"] or 0),
                float(row["variacion"] or 0),
                float(row["participacion_pct"] or 0),
                float(row.get("objetivo_categoria") or 0),
                float(row["uso_objetivo_pct"] or 0) if row.get("uso_objetivo_pct") is not None else None,
                row.get("objetivo_estado_label", ""),
            ]
        )

    ws_solicitudes = wb.create_sheet(title="Solicitudes")
    ws_solicitudes.append(
        [
            "Folio",
            "Area",
            "Solicitante",
            "Insumo",
            "Proveedor sugerido",
            "Cantidad",
            "Costo unitario",
            "Presupuesto",
            "Fecha requerida",
            "Estatus",
        ]
    )
    for s in solicitudes:
        ws_solicitudes.append(
            [
                s.folio,
                s.area,
                s.solicitante,
                s.insumo.nombre,
                s.proveedor_sugerido.nombre if s.proveedor_sugerido_id else "",
                float(s.cantidad or 0),
                float(s.costo_unitario or 0),
                float(s.presupuesto_estimado or 0),
                s.fecha_requerida.isoformat() if s.fecha_requerida else "",
                s.get_estatus_display(),
            ]
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="compras_consolidado_{now_str}.xlsx"'
    return response


def _export_tablero_proveedor_csv(
    provider_dashboard: dict,
    periodo_label: str,
    source_filter: str,
    plan_filter: str,
    categoria_filter: str,
) -> HttpResponse:
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="compras_tablero_proveedor_{now_str}.csv"'
    writer = csv.writer(response)
    writer.writerow(["TABLERO PROVEEDOR - COMPRAS"])
    writer.writerow(["Periodo activo", periodo_label])
    writer.writerow(["Filtro origen", source_filter])
    writer.writerow(["Filtro plan", plan_filter or "-"])
    writer.writerow(["Filtro categoria", categoria_filter or "-"])
    writer.writerow([])

    writer.writerow(["TOP DESVIACIONES (PERIODO ACTIVO)"])
    writer.writerow(
        [
            "Proveedor",
            "Estimado",
            "Ejecutado",
            "Variacion",
            "Participacion %",
            "Objetivo proveedor",
            "% Uso objetivo proveedor",
            "Estado objetivo",
        ]
    )
    for row in provider_dashboard["top_desviaciones"]:
        writer.writerow(
            [
                row["proveedor"],
                row["estimado"],
                row["ejecutado"],
                row["variacion"],
                round(float(row["participacion_pct"] or 0), 2),
                row.get("objetivo_proveedor", Decimal("0")),
                round(float(row["uso_objetivo_pct"] or 0), 2) if row.get("uso_objetivo_pct") is not None else "",
                row.get("objetivo_estado_label", ""),
            ]
        )
    writer.writerow([])

    writer.writerow(["TENDENCIA 6 MESES (TOP PROVEEDORES)"])
    writer.writerow(["Proveedor", "Mes", "Estimado", "Ejecutado", "Variacion"])
    for row in provider_dashboard["trend_rows"]:
        writer.writerow(
            [
                row["proveedor"],
                row["mes"],
                row["estimado"],
                row["ejecutado"],
                row["variacion"],
            ]
        )

    return response


def _export_tablero_proveedor_xlsx(
    provider_dashboard: dict,
    periodo_label: str,
    source_filter: str,
    plan_filter: str,
    categoria_filter: str,
) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Top desviaciones"
    ws.append(["TABLERO PROVEEDOR - COMPRAS"])
    ws.append(["Periodo activo", periodo_label])
    ws.append(["Filtro origen", source_filter])
    ws.append(["Filtro plan", plan_filter or "-"])
    ws.append(["Filtro categoria", categoria_filter or "-"])
    ws.append([])
    ws.append(
        [
            "Proveedor",
            "Estimado",
            "Ejecutado",
            "Variacion",
            "Participacion %",
            "Objetivo proveedor",
            "% Uso objetivo proveedor",
            "Estado objetivo",
        ]
    )
    for row in provider_dashboard["top_desviaciones"]:
        ws.append(
            [
                row["proveedor"],
                float(row["estimado"] or 0),
                float(row["ejecutado"] or 0),
                float(row["variacion"] or 0),
                float(row["participacion_pct"] or 0),
                float(row.get("objetivo_proveedor") or 0),
                float(row["uso_objetivo_pct"] or 0) if row.get("uso_objetivo_pct") is not None else None,
                row.get("objetivo_estado_label", ""),
            ]
        )

    ws2 = wb.create_sheet(title="Tendencia 6m")
    ws2.append(["Proveedor", "Mes", "Estimado", "Ejecutado", "Variacion"])
    for row in provider_dashboard["trend_rows"]:
        ws2.append(
            [
                row["proveedor"],
                row["mes"],
                float(row["estimado"] or 0),
                float(row["ejecutado"] or 0),
                float(row["variacion"] or 0),
            ]
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="compras_tablero_proveedor_{now_str}.xlsx"'
    return response


def _export_tablero_categoria_csv(
    category_dashboard: dict,
    periodo_label: str,
    source_filter: str,
    plan_filter: str,
    categoria_filter: str,
) -> HttpResponse:
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="compras_tablero_categoria_{now_str}.csv"'
    writer = csv.writer(response)
    writer.writerow(["TABLERO CATEGORIA - COMPRAS"])
    writer.writerow(["Periodo activo", periodo_label])
    writer.writerow(["Filtro origen", source_filter])
    writer.writerow(["Filtro plan", plan_filter or "-"])
    writer.writerow(["Filtro categoria", categoria_filter or "-"])
    writer.writerow([])

    writer.writerow(["TOP DESVIACIONES (PERIODO ACTIVO)"])
    writer.writerow(
        [
            "Categoria",
            "Estimado",
            "Ejecutado",
            "Variacion",
            "Participacion %",
            "Objetivo categoria",
            "% Uso objetivo categoria",
            "Estado objetivo",
        ]
    )
    for row in category_dashboard["top_desviaciones"]:
        writer.writerow(
            [
                row["categoria"],
                row["estimado"],
                row["ejecutado"],
                row["variacion"],
                round(float(row["participacion_pct"] or 0), 2),
                row.get("objetivo_categoria", Decimal("0")),
                round(float(row["uso_objetivo_pct"] or 0), 2) if row.get("uso_objetivo_pct") is not None else "",
                row.get("objetivo_estado_label", ""),
            ]
        )
    writer.writerow([])

    writer.writerow(["TENDENCIA 6 MESES (TOP CATEGORIAS)"])
    writer.writerow(["Categoria", "Mes", "Estimado", "Ejecutado", "Variacion"])
    for row in category_dashboard["trend_rows"]:
        writer.writerow(
            [
                row["categoria"],
                row["mes"],
                row["estimado"],
                row["ejecutado"],
                row["variacion"],
            ]
        )
    return response


def _export_tablero_categoria_xlsx(
    category_dashboard: dict,
    periodo_label: str,
    source_filter: str,
    plan_filter: str,
    categoria_filter: str,
) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Top desviaciones cat"
    ws.append(["TABLERO CATEGORIA - COMPRAS"])
    ws.append(["Periodo activo", periodo_label])
    ws.append(["Filtro origen", source_filter])
    ws.append(["Filtro plan", plan_filter or "-"])
    ws.append(["Filtro categoria", categoria_filter or "-"])
    ws.append([])
    ws.append(
        [
            "Categoria",
            "Estimado",
            "Ejecutado",
            "Variacion",
            "Participacion %",
            "Objetivo categoria",
            "% Uso objetivo categoria",
            "Estado objetivo",
        ]
    )
    for row in category_dashboard["top_desviaciones"]:
        ws.append(
            [
                row["categoria"],
                float(row["estimado"] or 0),
                float(row["ejecutado"] or 0),
                float(row["variacion"] or 0),
                float(row["participacion_pct"] or 0),
                float(row.get("objetivo_categoria") or 0),
                float(row["uso_objetivo_pct"] or 0) if row.get("uso_objetivo_pct") is not None else None,
                row.get("objetivo_estado_label", ""),
            ]
        )

    ws2 = wb.create_sheet(title="Tendencia 6m cat")
    ws2.append(["Categoria", "Mes", "Estimado", "Ejecutado", "Variacion"])
    for row in category_dashboard["trend_rows"]:
        ws2.append(
            [
                row["categoria"],
                row["mes"],
                float(row["estimado"] or 0),
                float(row["ejecutado"] or 0),
                float(row["variacion"] or 0),
            ]
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="compras_tablero_categoria_{now_str}.xlsx"'
    return response


def _export_consumo_plan_csv(
    consumo_dashboard: dict,
    periodo_label: str,
    source_filter: str,
    plan_filter: str,
    categoria_filter: str,
    consumo_ref_filter: str,
) -> HttpResponse:
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="compras_consumo_vs_plan_{now_str}.csv"'
    writer = csv.writer(response)
    writer.writerow(["CONSUMO REAL VS PLAN DE PRODUCCION"])
    writer.writerow(["Periodo activo", periodo_label])
    writer.writerow(["Rango considerado", f"{consumo_dashboard['period_start']} a {consumo_dashboard['period_end']}"])
    writer.writerow(["Filtro origen", source_filter])
    writer.writerow(["Filtro plan", plan_filter or "-"])
    writer.writerow(["Filtro categoria", categoria_filter or "-"])
    writer.writerow(
        [
            "Filtro consumo ref",
            "Solo referencia plan" if consumo_ref_filter == "plan_ref" else "Todos",
        ]
    )
    writer.writerow([])
    writer.writerow(["Totales"])
    writer.writerow(["Plan qty", consumo_dashboard["totals"]["plan_qty_total"]])
    writer.writerow(["Real qty", consumo_dashboard["totals"]["consumo_real_qty_total"]])
    writer.writerow(["Plan costo", consumo_dashboard["totals"]["plan_cost_total"]])
    writer.writerow(["Real costo", consumo_dashboard["totals"]["consumo_real_cost_total"]])
    writer.writerow(["Variacion costo", consumo_dashboard["totals"]["variacion_cost_total"]])
    writer.writerow(["Cobertura %", consumo_dashboard["totals"]["cobertura_pct"] or ""])
    writer.writerow([])
    writer.writerow(
        [
            "Insumo",
            "Categoria",
            "Unidad",
            "Cantidad plan",
            "Cantidad real",
            "Variacion qty",
            "Costo unitario",
            "Costo plan",
            "Costo real",
            "Variacion costo",
            "Consumo %",
            "Estado",
            "Semaforo",
            "Alerta costo",
        ]
    )
    for row in consumo_dashboard["rows"]:
        writer.writerow(
            [
                row["insumo"],
                row["categoria"],
                row["unidad"],
                row["cantidad_plan"],
                row["cantidad_real"],
                row["variacion_qty"],
                row["costo_unitario"],
                row["costo_plan"],
                row["costo_real"],
                row["variacion_cost"],
                row["consumo_pct"] if row["consumo_pct"] is not None else "",
                row["estado"],
                row["semaforo"],
                "SI" if row["sin_costo"] else "",
            ]
        )
    return response


def _export_consumo_plan_xlsx(
    consumo_dashboard: dict,
    periodo_label: str,
    source_filter: str,
    plan_filter: str,
    categoria_filter: str,
    consumo_ref_filter: str,
) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Consumo vs plan"
    ws.append(["CONSUMO REAL VS PLAN DE PRODUCCION"])
    ws.append(["Periodo activo", periodo_label])
    ws.append(["Rango considerado", f"{consumo_dashboard['period_start']} a {consumo_dashboard['period_end']}"])
    ws.append(["Filtro origen", source_filter])
    ws.append(["Filtro plan", plan_filter or "-"])
    ws.append(["Filtro categoria", categoria_filter or "-"])
    ws.append(["Filtro consumo ref", "Solo referencia plan" if consumo_ref_filter == "plan_ref" else "Todos"])
    ws.append([])
    ws.append(["Plan qty", float(consumo_dashboard["totals"]["plan_qty_total"] or 0)])
    ws.append(["Real qty", float(consumo_dashboard["totals"]["consumo_real_qty_total"] or 0)])
    ws.append(["Plan costo", float(consumo_dashboard["totals"]["plan_cost_total"] or 0)])
    ws.append(["Real costo", float(consumo_dashboard["totals"]["consumo_real_cost_total"] or 0)])
    ws.append(["Variacion costo", float(consumo_dashboard["totals"]["variacion_cost_total"] or 0)])
    ws.append(["Cobertura %", float(consumo_dashboard["totals"]["cobertura_pct"] or 0)])
    ws.append([])
    ws.append(
        [
            "Insumo",
            "Categoria",
            "Unidad",
            "Cantidad plan",
            "Cantidad real",
            "Variacion qty",
            "Costo unitario",
            "Costo plan",
            "Costo real",
            "Variacion costo",
            "Consumo %",
            "Estado",
            "Semaforo",
            "Alerta costo",
        ]
    )
    for row in consumo_dashboard["rows"]:
        ws.append(
            [
                row["insumo"],
                row["categoria"],
                row["unidad"],
                float(row["cantidad_plan"] or 0),
                float(row["cantidad_real"] or 0),
                float(row["variacion_qty"] or 0),
                float(row["costo_unitario"] or 0),
                float(row["costo_plan"] or 0),
                float(row["costo_real"] or 0),
                float(row["variacion_cost"] or 0),
                float(row["consumo_pct"] or 0) if row["consumo_pct"] is not None else None,
                row["estado"],
                row["semaforo"],
                "SI" if row["sin_costo"] else "",
            ]
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="compras_consumo_vs_plan_{now_str}.xlsx"'
    return response


def _filtered_solicitudes(
    source_filter_raw: str,
    plan_filter_raw: str,
    categoria_filter_raw: str,
    reabasto_filter_raw: str,
    estatus_filter_raw: str,
    workflow_action_raw: str,
    blocker_key_raw: str,
    periodo_tipo_raw: str,
    periodo_mes_raw: str,
):
    source_filter = (source_filter_raw or "all").lower()
    if source_filter not in {"all", "manual", "plan"}:
        source_filter = "all"
    plan_filter = (plan_filter_raw or "").strip()
    categoria_filter = _sanitize_categoria_filter(categoria_filter_raw)
    estatus_filter = (estatus_filter_raw or "all").strip().upper() or "ALL"
    periodo_tipo, periodo_mes, periodo_label = _parse_period_filters(periodo_tipo_raw, periodo_mes_raw)

    solicitudes_qs = SolicitudCompra.objects.select_related("insumo", "insumo__unidad_base", "proveedor_sugerido").all()
    if source_filter == "plan":
        solicitudes_qs = solicitudes_qs.filter(area__startswith="PLAN_PRODUCCION:")
    elif source_filter == "manual":
        solicitudes_qs = solicitudes_qs.exclude(area__startswith="PLAN_PRODUCCION:")

    if plan_filter:
        solicitudes_qs = solicitudes_qs.filter(area=f"PLAN_PRODUCCION:{plan_filter}")
    solicitudes_qs = _filter_solicitudes_by_categoria(solicitudes_qs, categoria_filter)

    if periodo_tipo != "all":
        year, month = periodo_mes.split("-")
        y = int(year)
        m = int(month)
        solicitudes_qs = solicitudes_qs.filter(fecha_requerida__year=y, fecha_requerida__month=m)
        if periodo_tipo == "q1":
            solicitudes_qs = solicitudes_qs.filter(fecha_requerida__day__lte=15)
        elif periodo_tipo == "q2":
            solicitudes_qs = solicitudes_qs.filter(fecha_requerida__day__gte=16)

    solicitudes = list(solicitudes_qs[:300])
    member_to_row, canonical_by_id = _canonical_catalog_maps()
    canonical_ids = {
        canonical_insumo_by_id(s.insumo_id).id
        for s in solicitudes
        if canonical_insumo_by_id(s.insumo_id)
    }
    canonical_member_ids = sorted(
        {
            member.id
            for canonical_id in canonical_ids
            for member in (canonical_by_id.get(canonical_id, {}) or {}).get("items", [])
        }
    )
    existencias_raw = {
        e.insumo_id: e
        for e in ExistenciaInsumo.objects.filter(insumo_id__in=canonical_member_ids)
    }

    plan_ids = set()
    for s in solicitudes:
        if (s.area or "").startswith("PLAN_PRODUCCION:"):
            _, _, maybe_id = s.area.partition(":")
            if maybe_id.isdigit():
                plan_ids.add(int(maybe_id))
    planes_map = {
        p.id: p
        for p in PlanProduccion.objects.filter(id__in=plan_ids)
    }
    latest_cost_by_insumo = _latest_cost_by_canonical_ids(canonical_ids, canonical_by_id)

    for s in solicitudes:
        canonical = canonical_insumo_by_id(s.insumo_id)
        row = member_to_row.get(s.insumo_id)
        member_ids = row["member_ids"] if row else []
        member_existencias = [existencias_raw[mid] for mid in member_ids if mid in existencias_raw]
        ex = existencias_raw.get(canonical.id) if canonical else None
        ex = ex or next((item for item in member_existencias if item), None)
        stock_actual = sum((item.stock_actual for item in member_existencias), Decimal("0"))
        punto_reorden = ex.punto_reorden if ex else Decimal("0")
        if stock_actual <= Decimal("0"):
            s.reabasto_nivel = "critico"
            s.reabasto_texto = "Sin stock"
        elif stock_actual < punto_reorden:
            s.reabasto_nivel = "bajo"
            s.reabasto_texto = "Bajo reorden"
        else:
            s.reabasto_nivel = "ok"
            s.reabasto_texto = "Stock suficiente"
        s.reabasto_detalle = f"Stock {stock_actual} / Reorden {punto_reorden}"
        s.__dict__.update(_source_context_from_scope(area=s.area, planes_map=planes_map))
        s.costo_unitario = latest_cost_by_insumo.get(canonical.id, Decimal("0")) if canonical else Decimal("0")
        s.presupuesto_estimado = (s.cantidad or Decimal("0")) * (s.costo_unitario or Decimal("0"))

    open_orders_by_solicitud = {}
    solicitud_ids = [s.id for s in solicitudes]
    if solicitud_ids:
        for orden in (
            OrdenCompra.objects.filter(solicitud_id__in=solicitud_ids)
            .exclude(estatus=OrdenCompra.STATUS_CERRADA)
            .order_by("-creado_en")
        ):
            open_orders_by_solicitud.setdefault(orden.solicitud_id, orden)

    for s in solicitudes:
        open_order = open_orders_by_solicitud.get(s.id)
        s.has_open_order = bool(open_order)
        s.open_order_id = open_order.id if open_order else None
        s.open_order_folio = open_order.folio if open_order else ""
        _enrich_solicitud_workflow(s)

    valid_statuses = {choice[0] for choice in SolicitudCompra.STATUS_CHOICES}
    if estatus_filter == "BLOCKED_ERP":
        solicitudes = [s for s in solicitudes if getattr(s, "has_workflow_blockers", False)]
    elif estatus_filter == "APPROVED_READY":
        solicitudes = [s for s in solicitudes if s.estatus == SolicitudCompra.STATUS_APROBADA and not s.has_open_order]
    elif estatus_filter == "APPROVED_WITH_OC":
        solicitudes = [s for s in solicitudes if s.estatus == SolicitudCompra.STATUS_APROBADA and s.has_open_order]
    elif estatus_filter in valid_statuses:
        solicitudes = [s for s in solicitudes if s.estatus == estatus_filter]
    else:
        estatus_filter = "ALL"

    reabasto_filter = (reabasto_filter_raw or "all").lower()
    if reabasto_filter in {"critico", "bajo", "ok"}:
        solicitudes = [s for s in solicitudes if s.reabasto_nivel == reabasto_filter]
    else:
        reabasto_filter = "all"
    workflow_action_filter = (workflow_action_raw or "all").strip().lower()
    valid_workflow_actions = {
        "all",
        "corregir_maestro",
        "enviar_revision",
        "aprobar_rechazar",
        "crear_oc",
        "seguimiento_oc",
    }
    if workflow_action_filter in valid_workflow_actions and workflow_action_filter != "all":
        solicitudes = [s for s in solicitudes if getattr(s, "workflow_action_code", "") == workflow_action_filter]
    else:
        workflow_action_filter = "all"
    blocker_key_filter = (blocker_key_raw or "all").strip().lower()
    valid_blocker_keys = {"all", "sin_costo", "sin_proveedor", "maestro_incompleto", "articulo_inactivo", "sin_catalogo", "no_canonico"}
    if blocker_key_filter in valid_blocker_keys and blocker_key_filter != "all":
        solicitudes = [s for s in solicitudes if _solicitud_has_blocker_key(s, blocker_key_filter)]
    else:
        blocker_key_filter = "all"

    plan_ids_all = set()
    for area_val in SolicitudCompra.objects.filter(area__startswith="PLAN_PRODUCCION:").values_list("area", flat=True).distinct():
        _, _, maybe_id = (area_val or "").partition(":")
        if maybe_id.isdigit():
            plan_ids_all.add(int(maybe_id))
    plan_options = list(PlanProduccion.objects.filter(id__in=plan_ids_all).order_by("-fecha_produccion", "-id")[:100])

    return (
        solicitudes,
        source_filter,
        plan_filter,
        categoria_filter,
        reabasto_filter,
        estatus_filter,
        workflow_action_filter,
        blocker_key_filter,
        plan_options,
        periodo_tipo,
        periodo_mes,
        periodo_label,
    )


def _document_release_gate_completion(rows: list[dict[str, object]]) -> int:
    total = sum(int(row.get("total") or 0) for row in rows)
    if total <= 0:
        return 0
    completed = sum(int(row.get("completed") or 0) for row in rows)
    return int(round((completed / total) * 100))


def _document_governance_rows(
    rows: list[dict[str, object]], *, owner: str = "Compras / Operación"
) -> list[dict[str, object]]:
    governance_rows: list[dict[str, object]] = []
    for row in rows:
        total = max(int(row.get("total") or 0), 1)
        completed = int(row.get("completed") or 0)
        governance_rows.append(
            {
                "front": row.get("title", "Frente"),
                "owner": owner,
                "blockers": int(row.get("open_count") or 0),
                "completion": int(round((completed / total) * 100)) if total else 0,
                "detail": row.get("detail", ""),
                "next_step": row.get("cta", "Abrir"),
                "url": row.get("url", reverse("compras:solicitudes")),
                "cta": "Abrir",
            }
        )
    return governance_rows


def _document_critical_path_rows(
    rows: list[dict[str, object]], *, owner: str = "Compras / Operación", fallback_url: str
) -> list[dict[str, object]]:
    ranked = sorted(
        rows,
        key=lambda row: (
            -int(row.get("open_count") or 0),
            int(row.get("completion") or 0),
            -int(row.get("total") or 0),
        ),
    )
    critical_rows: list[dict[str, object]] = []
    for index, row in enumerate(ranked[:4], start=1):
        critical_rows.append(
            {
                "rank": f"R{index}",
                "title": row.get("title", "Tramo documental"),
                "owner": owner,
                "status": row.get("status", "Sin estado"),
                "count": int(row.get("open_count") or 0),
                "completion": int(row.get("completion") or 0),
                "depends_on": row.get("depends_on", "Inicio del flujo"),
                "dependency_status": row.get("dependency_status", "Sin dependencia registrada"),
                "detail": row.get("detail", ""),
                "next_step": row.get("cta", "Abrir"),
                "url": row.get("url", fallback_url),
                "cta": "Abrir",
            }
        )
    return critical_rows


def _document_executive_radar_rows(
    rows: list[dict[str, object]], *, owner: str = "Compras / Operación", fallback_url: str
) -> list[dict[str, object]]:
    radar_rows: list[dict[str, object]] = []
    for row in rows[:4]:
        completion = int(row.get("completion") or 0)
        blockers = int(row.get("open_count") or 0)
        if blockers <= 0 and completion >= 100:
            tone = "success"
            status = "Controlado"
            dominant_blocker = "Sin bloqueo activo"
        elif completion >= 50:
            tone = "warning"
            status = "En seguimiento"
            dominant_blocker = row.get("detail", "") or "Seguimiento documental pendiente"
        else:
            tone = "danger"
            status = "Con bloqueo"
            dominant_blocker = row.get("detail", "") or "Bloqueo documental abierto"
        radar_rows.append(
            {
                "phase": row.get("title", "Fase documental"),
                "owner": owner,
                "status": status,
                "tone": tone,
                "blockers": blockers,
                "progress_pct": completion,
                "dominant_blocker": dominant_blocker,
                "depends_on": row.get("depends_on", "Inicio del flujo"),
                "dependency_status": row.get("dependency_status", "Sin dependencia registrada"),
                "next_step": row.get("cta", "Abrir"),
                "url": row.get("url", fallback_url),
                "cta": "Abrir",
            }
        )
    return radar_rows


def _document_trunk_handoff_rows(
    rows: list[dict[str, object]], *, owner: str = "Compras / Operación", fallback_url: str
) -> list[dict[str, object]]:
    labels = [
        ("Solicitudes / BOM", "Artículo liberado y necesidad documental clara"),
        ("Órdenes documentales", "Solicitud aprobada y convertida a compra operativa"),
        ("Recepciones / Inventario", "Recepción aplicada y abastecimiento conciliado"),
    ]
    trunk_rows: list[dict[str, object]] = []
    for index, (label, depends_on) in enumerate(labels):
        row = rows[index] if index < len(rows) else {}
        total = max(int(row.get("total") or 0), 1)
        completed = int(row.get("completed") or 0)
        completion = int(round((completed / total) * 100)) if total else 0
        trunk_rows.append(
            {
                "label": label,
                "owner": owner,
                "status": "Listo para operar" if int(row.get("open_count") or 0) == 0 and completion >= 100 else "Bloqueado",
                "tone": "success" if int(row.get("open_count") or 0) == 0 and completion >= 100 else "warning",
                "blockers": int(row.get("open_count") or 0),
                "completion": completion,
                "depends_on": depends_on,
                "exit_criteria": row.get("detail") or "La etapa debe quedar cerrada para alimentar el siguiente tramo.",
                "detail": row.get("detail") or "Sin detalle documental.",
                "next_step": row.get("cta") or "Abrir",
                "url": row.get("url") or fallback_url,
                "cta": "Abrir",
            }
        )
    return trunk_rows


def _document_command_center(
    release_gate_rows: list[dict[str, object]],
    *,
    owner: str,
    plan_scope_context: dict | None = None,
    fallback_url: str,
    fallback_cta: str,
) -> dict[str, object]:
    blockers = sum(int(row.get("open_count") or 0) for row in release_gate_rows)
    if plan_scope_context:
        status = plan_scope_context.get("summary_label", "Controlado")
        tone = (
            "danger"
            if blockers
            else "warning"
            if plan_scope_context.get("document_progress_pct", 0) < 100
            else "success"
        )
        next_step = plan_scope_context.get("next_action", {}).get("detail") or "Continuar el cierre documental del plan."
        url = plan_scope_context.get("next_action", {}).get("url") or fallback_url
        cta = plan_scope_context.get("next_action", {}).get("label") or fallback_cta
    else:
        completion = _document_release_gate_completion(release_gate_rows)
        if blockers:
            status = "Crítico"
            tone = "danger"
            next_step = "Resolver bloqueos documentales antes de avanzar a la siguiente etapa."
        elif completion < 100:
            status = "Seguimiento"
            tone = "warning"
            next_step = "Cerrar las validaciones pendientes del flujo documental."
        else:
            status = "Controlado"
            tone = "success"
            next_step = "Mantener el flujo documental sin bloqueos."
        url = fallback_url
        cta = fallback_cta
    return {
        "owner": owner,
        "status": status,
        "tone": tone,
        "blockers": blockers,
        "next_step": next_step,
        "url": url,
        "cta": cta,
    }


def _ordenes_release_gate_rows(ordenes: list[OrdenCompra]) -> list[dict[str, object]]:
    total = max(len(ordenes), 1)
    article_open = sum(1 for orden in ordenes if getattr(orden, "enterprise_master_blocker_details", []))
    article_completed = max(len(ordenes) - article_open, 0)
    order_open = sum(
        1
        for orden in ordenes
        if orden.estatus == OrdenCompra.STATUS_BORRADOR or getattr(orden, "has_workflow_blockers", False)
    )
    order_completed = max(len(ordenes) - order_open, 0)
    reception_open = sum(1 for orden in ordenes if orden.estatus != OrdenCompra.STATUS_CERRADA)
    reception_completed = max(len(ordenes) - reception_open, 0)
    return [
        {
            "step": "01",
            "title": "Articulo liberado",
            "detail": "La orden parte de articulos con maestro, costo y referencia documental listos.",
            "completed": article_completed,
            "open_count": article_open,
            "total": total,
            "tone": "success" if article_open == 0 else "warning",
            "url": reverse("compras:ordenes"),
            "cta": "Abrir ordenes",
        },
        {
            "step": "02",
            "title": "Orden emitida",
            "detail": "La orden ya fue emitida o confirmada y puede avanzar al proceso de recepcion.",
            "completed": order_completed,
            "open_count": order_open,
            "total": total,
            "tone": "success" if order_open == 0 else "warning",
            "url": reverse("compras:ordenes"),
            "cta": "Revisar ordenes",
        },
        {
            "step": "03",
            "title": "Recepcion cerrada",
            "detail": "La cadena documental cierra cuando la orden ya fue recibida y conciliada.",
            "completed": reception_completed,
            "open_count": reception_open,
            "total": total,
            "tone": "success" if reception_open == 0 else "warning",
            "url": reverse("compras:recepciones"),
            "cta": "Abrir recepciones",
        },
    ]


def _recepciones_release_gate_rows(recepciones: list[RecepcionCompra]) -> list[dict[str, object]]:
    total = max(len(recepciones), 1)
    article_open = sum(1 for recepcion in recepciones if getattr(recepcion, "enterprise_master_blocker_details", []))
    article_completed = max(len(recepciones) - article_open, 0)
    order_open = sum(
        1
        for recepcion in recepciones
        if recepcion.orden.estatus == OrdenCompra.STATUS_BORRADOR or getattr(recepcion, "has_workflow_blockers", False)
    )
    order_completed = max(len(recepciones) - order_open, 0)
    reception_open = sum(1 for recepcion in recepciones if recepcion.estatus != RecepcionCompra.STATUS_CERRADA)
    reception_completed = max(len(recepciones) - reception_open, 0)
    return [
        {
            "step": "01",
            "title": "Articulo liberado",
            "detail": "La recepcion opera sobre articulos completos y listos en maestro ERP.",
            "completed": article_completed,
            "open_count": article_open,
            "total": total,
            "tone": "success" if article_open == 0 else "warning",
            "url": reverse("compras:recepciones"),
            "cta": "Abrir recepciones",
        },
        {
            "step": "02",
            "title": "Orden emitida",
            "detail": "La orden origen ya esta formalizada y puede registrar entrada documental.",
            "completed": order_completed,
            "open_count": order_open,
            "total": total,
            "tone": "success" if order_open == 0 else "warning",
            "url": reverse("compras:ordenes"),
            "cta": "Abrir ordenes",
        },
        {
            "step": "03",
            "title": "Recepcion cerrada",
            "detail": "El abastecimiento queda liberado cuando la recepcion ya cerro y aplico inventario.",
            "completed": reception_completed,
            "open_count": reception_open,
            "total": total,
            "tone": "success" if reception_open == 0 else "warning",
            "url": reverse("compras:recepciones"),
            "cta": "Revisar recepciones",
        },
    ]


@login_required
def solicitudes(request: HttpRequest) -> HttpResponse:
    if not can_view_compras(request.user):
        raise PermissionDenied("No tienes permisos para ver Compras.")

    if request.method == "POST":
        if not can_manage_compras(request.user):
            raise PermissionDenied("No tienes permisos para crear solicitudes.")
        locked_plan_scope = _locked_plan_scope_from_request(request, current_view="solicitudes")
        if (request.POST.get("source") or "").strip().lower() == "plan" and not locked_plan_scope:
            messages.error(request, "No se pudo resolver el plan activo para la solicitud.")
            return _redirect_scoped_list("compras:solicitudes", request)
        insumo_id = request.POST.get("insumo_id")
        if insumo_id:
            insumo = canonical_insumo_by_id(insumo_id)
            if not insumo:
                messages.error(request, "El insumo seleccionado no es válido.")
                return _redirect_scoped_list("compras:solicitudes", request)
            solicitud = SolicitudCompra.objects.create(
                area=(
                    locked_plan_scope["plan_scope"]
                    if locked_plan_scope
                    else (request.POST.get("area", "General").strip() or "General")
                ),
                solicitante=request.POST.get("solicitante", request.user.username).strip() or request.user.username,
                insumo=insumo,
                proveedor_sugerido=insumo.proveedor_principal,
                cantidad=_to_decimal(request.POST.get("cantidad"), "1"),
                fecha_requerida=request.POST.get("fecha_requerida") or date.today(),
                estatus=request.POST.get("estatus") or SolicitudCompra.STATUS_BORRADOR,
            )
            log_event(
                request.user,
                "CREATE",
                "compras.SolicitudCompra",
                solicitud.id,
                {"folio": solicitud.folio, "estatus": solicitud.estatus},
            )
        return _redirect_scoped_list("compras:solicitudes", request)

    (
        solicitudes,
        source_filter,
        plan_filter,
        categoria_filter,
        reabasto_filter,
        estatus_filter,
        workflow_action_filter,
        blocker_key_filter,
        plan_options,
        periodo_tipo,
        periodo_mes,
        periodo_label,
    ) = _filtered_solicitudes(
        request.GET.get("source"),
        request.GET.get("plan_id"),
        request.GET.get("categoria"),
        request.GET.get("reabasto"),
        request.GET.get("estatus"),
        request.GET.get("workflow_action"),
        request.GET.get("blocker_key"),
        request.GET.get("periodo_tipo"),
        request.GET.get("periodo_mes"),
    )
    closure_key_raw = (request.GET.get("closure_key") or "all").strip().lower()
    handoff_key_raw = (request.GET.get("handoff_key") or "all").strip().lower()
    closure_key_filter = closure_key_raw
    handoff_key_filter = handoff_key_raw
    master_class_filter = (request.GET.get("master_class") or "all").strip()
    master_missing_filter = (request.GET.get("master_missing") or "all").strip()
    for solicitud in solicitudes:
        solicitud.enterprise_master_blocker_details = _enterprise_blocker_details_for_solicitud(solicitud)
    valid_closure_keys = {"all", "solicitudes_liberadas"}
    if closure_key_filter in valid_closure_keys and closure_key_filter != "all":
        solicitudes = [
            solicitud
            for solicitud in solicitudes
            if getattr(solicitud, "has_workflow_blockers", False)
            or solicitud.estatus in {SolicitudCompra.STATUS_BORRADOR, SolicitudCompra.STATUS_EN_REVISION}
        ]
    else:
        closure_key_filter = "all"
    valid_handoff_keys = {"all", "solicitud_orden"}
    if handoff_key_filter in valid_handoff_keys and handoff_key_filter != "all":
        solicitudes = [
            solicitud
            for solicitud in solicitudes
            if getattr(solicitud, "has_workflow_blockers", False)
            or (
                solicitud.estatus == SolicitudCompra.STATUS_APROBADA
                and not getattr(solicitud, "has_open_order", False)
            )
        ]
    else:
        handoff_key_filter = "all"
    solicitudes = _filter_documents_by_master_blockers(
        solicitudes,
        master_class_filter,
        master_missing_filter,
    )
    consumo_ref_filter = _sanitize_consumo_ref_filter(request.GET.get("consumo_ref"))
    budget_ctx = _build_budget_context(
        solicitudes,
        source_filter,
        plan_filter,
        categoria_filter,
        periodo_tipo,
        periodo_mes,
    )
    provider_dashboard = _build_provider_dashboard(
        periodo_mes,
        source_filter,
        plan_filter,
        categoria_filter,
        budget_ctx["presupuesto_rows_proveedor"],
    )
    category_dashboard = _build_category_dashboard(
        periodo_mes,
        source_filter,
        plan_filter,
        categoria_filter,
        budget_ctx.get("presupuesto_rows_categoria", []),
    )
    consumo_dashboard = _build_consumo_vs_plan_dashboard(
        periodo_tipo,
        periodo_mes,
        source_filter,
        plan_filter,
        categoria_filter,
        consumo_ref_filter,
    )
    total_presupuesto = budget_ctx["presupuesto_estimado_total"]
    import_preview = _build_import_preview_context(request.session.get(IMPORT_PREVIEW_SESSION_KEY))

    export_format = (request.GET.get("export") or "").lower()
    if export_format == "csv":
        return _export_solicitudes_csv(
            solicitudes,
            source_filter,
            plan_filter,
            categoria_filter,
            reabasto_filter,
            periodo_tipo,
            periodo_mes,
            periodo_label,
        )
    if export_format == "consolidado_csv":
        return _export_consolidado_csv(
            solicitudes,
            source_filter,
            plan_filter,
            categoria_filter,
            reabasto_filter,
            periodo_label,
            budget_ctx,
        )
    if export_format == "consolidado_xlsx":
        return _export_consolidado_xlsx(
            solicitudes,
            source_filter,
            plan_filter,
            categoria_filter,
            reabasto_filter,
            periodo_label,
            budget_ctx,
        )
    if export_format == "proveedor_csv":
        return _export_tablero_proveedor_csv(
            provider_dashboard,
            periodo_label,
            source_filter,
            plan_filter,
            categoria_filter,
        )
    if export_format == "proveedor_xlsx":
        return _export_tablero_proveedor_xlsx(
            provider_dashboard,
            periodo_label,
            source_filter,
            plan_filter,
            categoria_filter,
        )
    if export_format == "categoria_csv":
        return _export_tablero_categoria_csv(
            category_dashboard,
            periodo_label,
            source_filter,
            plan_filter,
            categoria_filter,
        )
    if export_format == "categoria_xlsx":
        return _export_tablero_categoria_xlsx(
            category_dashboard,
            periodo_label,
            source_filter,
            plan_filter,
            categoria_filter,
        )
    if export_format == "consumo_plan_csv":
        return _export_consumo_plan_csv(
            consumo_dashboard,
            periodo_label,
            source_filter,
            plan_filter,
            categoria_filter,
            consumo_ref_filter,
        )
    if export_format == "consumo_plan_xlsx":
        return _export_consumo_plan_xlsx(
            consumo_dashboard,
            periodo_label,
            source_filter,
            plan_filter,
            categoria_filter,
            consumo_ref_filter,
        )
    if export_format == "import_preview_csv":
        if not import_preview:
            messages.warning(request, "No hay vista previa de importación activa para exportar.")
            return redirect("compras:solicitudes")
        return _export_import_preview_csv(import_preview)
    if export_format == "import_preview_xlsx":
        if not import_preview:
            messages.warning(request, "No hay vista previa de importación activa para exportar.")
            return redirect("compras:solicitudes")
        return _export_import_preview_xlsx(import_preview)
    if export_format == "xlsx":
        return _export_solicitudes_xlsx(
            solicitudes,
            source_filter,
            plan_filter,
            categoria_filter,
            reabasto_filter,
            periodo_tipo,
            periodo_mes,
            periodo_label,
        )

    query_without_export = request.GET.copy()
    query_without_export.pop("export", None)
    query_without_estatus = query_without_export.copy()
    query_without_estatus.pop("estatus", None)
    query_without_workflow = query_without_export.copy()
    query_without_workflow.pop("workflow_action", None)
    query_without_blocker = query_without_export.copy()
    query_without_blocker.pop("blocker_key", None)
    query_without_closure = query_without_export.copy()
    query_without_closure.pop("closure_key", None)
    query_without_handoff = query_without_export.copy()
    query_without_handoff.pop("handoff_key", None)
    plan_scope_context = _build_plan_scope_context(
        source_filter=source_filter,
        plan_filter=plan_filter,
        current_view="solicitudes",
        closure_key_filter=closure_key_raw,
        handoff_key_filter=handoff_key_raw,
        master_class_filter=master_class_filter,
        master_missing_filter=master_missing_filter,
        session=request.session,
    )

    workflow_summary = _solicitudes_workflow_summary(solicitudes)
    enterprise_board = _solicitudes_enterprise_board(solicitudes)
    supply_model_rows = _solicitudes_supply_model_rows(solicitudes)
    release_gate_rows = [
        {
            "step": "01",
            "title": "Artículo liberado",
            "detail": "El artículo ya tiene maestro, costo y referencia válidos para entrar al flujo documental.",
            "completed": sum(1 for s in solicitudes if not getattr(s, "is_enterprise_blocked", False)),
            "open_count": sum(1 for s in solicitudes if getattr(s, "is_enterprise_blocked", False)),
            "total": max(len(solicitudes), 1),
            "tone": "success"
            if not any(getattr(s, "is_enterprise_blocked", False) for s in solicitudes)
            else "warning",
            "url": reverse("compras:solicitudes"),
            "cta": "Abrir solicitudes",
        },
        {
            "step": "02",
            "title": "Orden emitida",
            "detail": "Las solicitudes aprobadas ya migraron a orden de compra sin bloqueo documental.",
            "completed": sum(int(row["closed"]) for row in supply_model_rows if row["step"] == "02"),
            "open_count": sum(int(row["pending"]) for row in supply_model_rows if row["step"] == "02"),
            "total": max(
                sum(
                    int(row["closed"]) + int(row["pending"])
                    for row in supply_model_rows
                    if row["step"] == "02"
                ),
                1,
            ),
            "tone": "success"
            if not any(int(row["pending"]) for row in supply_model_rows if row["step"] == "02")
            else "warning",
            "url": reverse("compras:ordenes"),
            "cta": "Abrir órdenes",
        },
        {
            "step": "03",
            "title": "Recepción cerrada",
            "detail": "Las órdenes ya recibidas y conciliadas permiten cerrar el abastecimiento del periodo.",
            "completed": sum(int(row["closed"]) for row in supply_model_rows if row["step"] == "03"),
            "open_count": sum(int(row["pending"]) for row in supply_model_rows if row["step"] == "03"),
            "total": max(
                sum(
                    int(row["closed"]) + int(row["pending"])
                    for row in supply_model_rows
                    if row["step"] == "03"
                ),
                1,
            ),
            "tone": "success"
            if not any(int(row["pending"]) for row in supply_model_rows if row["step"] == "03")
            else "warning",
            "url": reverse("compras:recepciones"),
            "cta": "Abrir recepciones",
        },
    ]
    context = {
        "solicitudes": solicitudes,
        "insumo_options": _build_insumo_options(),
        "proveedor_options": list(Proveedor.objects.filter(activo=True).only("id", "nombre").order_by("nombre")),
        "categoria_options": [
            c
            for c in sorted(
                {
                    *[
                        " ".join((x or "").strip().split())
                        for x in Insumo.objects.filter(activo=True)
                        .exclude(categoria="")
                        .values_list("categoria", flat=True)
                    ],
                    "Masa",
                    "Volumen",
                    "Pieza",
                }
            )
            if c
        ],
        "status_choices": SolicitudCompra.STATUS_CHOICES,
        "can_manage_compras": can_manage_compras(request.user),
        "reabasto_filter": reabasto_filter,
        "estatus_filter": estatus_filter,
        "workflow_action_filter": workflow_action_filter,
        "blocker_key_filter": blocker_key_filter,
        "closure_key_filter": closure_key_filter,
        "handoff_key_filter": handoff_key_filter,
        "source_filter": source_filter,
        "plan_filter": plan_filter,
        "categoria_filter": categoria_filter,
        "plan_options": plan_options,
        "periodo_tipo": periodo_tipo,
        "periodo_mes": periodo_mes,
        "periodo_label": periodo_label,
        "consumo_ref_filter": consumo_ref_filter,
        "master_class_filter": master_class_filter,
        "master_missing_filter": master_missing_filter,
        "master_class_choices": [
            ("all", "Todas las clases"),
            (Insumo.TIPO_MATERIA_PRIMA, "Materia prima"),
            (Insumo.TIPO_INTERNO, "Insumo interno"),
            (Insumo.TIPO_EMPAQUE, "Empaque"),
            ("unknown", "Sin catálogo"),
        ],
        "master_missing_choices": [
            ("all", "Todos los faltantes"),
            ("unidad", "Unidad base"),
            ("proveedor", "Proveedor principal"),
            ("categoria", "Categoría"),
            ("codigo_point", "Código Point"),
            ("other", "Otros"),
        ],
        "total_presupuesto": total_presupuesto,
        "current_query": query_without_export.urlencode(),
        "current_query_without_estatus": query_without_estatus.urlencode(),
        "current_query_without_workflow": query_without_workflow.urlencode(),
        "current_query_without_blocker": query_without_blocker.urlencode(),
        "current_query_without_closure": query_without_closure.urlencode(),
        "current_query_without_handoff": query_without_handoff.urlencode(),
        "plan_scope_context": plan_scope_context,
        "import_preview": import_preview,
        "presupuesto_historial": _build_budget_history(periodo_mes, source_filter, plan_filter, categoria_filter),
        "provider_dashboard": provider_dashboard,
        "category_dashboard": category_dashboard,
        "consumo_dashboard": consumo_dashboard,
        "workflow_summary": workflow_summary,
        "enterprise_board": enterprise_board,
        "supply_model_rows": supply_model_rows,
        "supply_model_completion": (
            sum(int(row["completion"]) for row in supply_model_rows) // len(supply_model_rows)
            if supply_model_rows
            else 0
        ),
        "release_gate_rows": release_gate_rows,
        "executive_radar_rows": _document_executive_radar_rows(
            release_gate_rows,
            owner="Compras / Abastecimiento",
            fallback_url=reverse("compras:solicitudes"),
        ),
        "trunk_handoff_rows": _document_trunk_handoff_rows(
            release_gate_rows,
            owner="Compras / Abastecimiento",
            fallback_url=reverse("compras:solicitudes"),
        ),
        "erp_governance_rows": _document_governance_rows(
            release_gate_rows,
            owner="Compras / Abastecimiento",
        ),
        "critical_path_rows": _document_critical_path_rows(
            release_gate_rows,
            owner="Compras / Abastecimiento",
            fallback_url=reverse("compras:solicitudes"),
        ),
        "erp_command_center": _document_command_center(
            release_gate_rows,
            owner="Compras / Abastecimiento",
            plan_scope_context=plan_scope_context,
            fallback_url=reverse("compras:solicitudes"),
            fallback_cta="Abrir solicitudes",
        ),
        "release_gate_completion": (
            round(
                (
                    sum(row["completed"] for row in release_gate_rows)
                    / sum(row["total"] for row in release_gate_rows)
                )
                * 100
            )
            if release_gate_rows and sum(row["total"] for row in release_gate_rows)
            else 0
        ),
        **budget_ctx,
    }
    return render(request, "compras/solicitudes.html", context)


@login_required
@require_POST
def guardar_presupuesto_periodo(request: HttpRequest) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para gestionar presupuesto.")

    periodo_tipo, periodo_mes, _ = _parse_period_filters(
        request.POST.get("periodo_tipo"),
        request.POST.get("periodo_mes"),
    )
    if periodo_tipo == "all":
        messages.error(request, "Selecciona periodo mensual o quincenal para guardar presupuesto.")
        return redirect("compras:solicitudes")

    monto_objetivo = _to_decimal(request.POST.get("monto_objetivo"), "0")
    if monto_objetivo < 0:
        monto_objetivo = Decimal("0")
    notas = (request.POST.get("notas") or "").strip()

    presupuesto, created = PresupuestoCompraPeriodo.objects.update_or_create(
        periodo_tipo=periodo_tipo,
        periodo_mes=periodo_mes,
        defaults={
            "monto_objetivo": monto_objetivo,
            "notas": notas,
            "actualizado_por": request.user,
        },
    )
    log_event(
        request.user,
        "CREATE" if created else "UPDATE",
        "compras.PresupuestoCompraPeriodo",
        presupuesto.id,
        {
            "periodo_tipo": periodo_tipo,
            "periodo_mes": periodo_mes,
            "monto_objetivo": str(monto_objetivo),
        },
    )
    messages.success(request, "Presupuesto del período actualizado.")

    params = {
        "source": (request.POST.get("source") or "all").strip() or "all",
        "plan_id": (request.POST.get("plan_id") or "").strip(),
        "categoria": _sanitize_categoria_filter(request.POST.get("categoria")),
        "reabasto": (request.POST.get("reabasto") or "all").strip() or "all",
        "consumo_ref": _sanitize_consumo_ref_filter(request.POST.get("consumo_ref")),
        "periodo_tipo": periodo_tipo,
        "periodo_mes": periodo_mes,
    }
    if not params["plan_id"]:
        params.pop("plan_id")
    if not params["categoria"]:
        params.pop("categoria")
    return redirect(f"{reverse('compras:solicitudes')}?{urlencode(params)}")


@login_required
@require_POST
def guardar_presupuesto_proveedor(request: HttpRequest) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para gestionar presupuesto por proveedor.")

    periodo_tipo, periodo_mes, _ = _parse_period_filters(
        request.POST.get("periodo_tipo"),
        request.POST.get("periodo_mes"),
    )
    if periodo_tipo == "all":
        messages.error(request, "Selecciona periodo mensual o quincenal para guardar objetivo por proveedor.")
        return redirect("compras:solicitudes")

    proveedor_id_raw = (request.POST.get("proveedor_id") or "").strip()
    if not proveedor_id_raw.isdigit():
        messages.error(request, "Selecciona un proveedor válido.")
        return redirect("compras:solicitudes")

    proveedor = get_object_or_404(Proveedor, pk=int(proveedor_id_raw), activo=True)
    monto_objetivo = _to_decimal(request.POST.get("monto_objetivo_proveedor"), "0")
    if monto_objetivo < 0:
        monto_objetivo = Decimal("0")
    notas = (request.POST.get("notas_proveedor") or "").strip()

    presupuesto_periodo, _ = PresupuestoCompraPeriodo.objects.get_or_create(
        periodo_tipo=periodo_tipo,
        periodo_mes=periodo_mes,
        defaults={"monto_objetivo": Decimal("0"), "actualizado_por": request.user},
    )
    objetivo_proveedor, created = PresupuestoCompraProveedor.objects.update_or_create(
        presupuesto_periodo=presupuesto_periodo,
        proveedor=proveedor,
        defaults={
            "monto_objetivo": monto_objetivo,
            "notas": notas,
            "actualizado_por": request.user,
        },
    )
    log_event(
        request.user,
        "CREATE" if created else "UPDATE",
        "compras.PresupuestoCompraProveedor",
        objetivo_proveedor.id,
        {
            "periodo_tipo": periodo_tipo,
            "periodo_mes": periodo_mes,
            "proveedor_id": proveedor.id,
            "proveedor_nombre": proveedor.nombre,
            "monto_objetivo": str(monto_objetivo),
        },
    )
    messages.success(request, f"Objetivo de proveedor actualizado: {proveedor.nombre}.")

    params = {
        "source": (request.POST.get("source") or "all").strip() or "all",
        "plan_id": (request.POST.get("plan_id") or "").strip(),
        "categoria": _sanitize_categoria_filter(request.POST.get("categoria")),
        "reabasto": (request.POST.get("reabasto") or "all").strip() or "all",
        "consumo_ref": _sanitize_consumo_ref_filter(request.POST.get("consumo_ref")),
        "periodo_tipo": periodo_tipo,
        "periodo_mes": periodo_mes,
    }
    if not params["plan_id"]:
        params.pop("plan_id")
    if not params["categoria"]:
        params.pop("categoria")
    return redirect(f"{reverse('compras:solicitudes')}?{urlencode(params)}")


@login_required
@require_POST
def guardar_presupuesto_categoria(request: HttpRequest) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para gestionar presupuesto por categoría.")

    periodo_tipo, periodo_mes, _ = _parse_period_filters(
        request.POST.get("periodo_tipo"),
        request.POST.get("periodo_mes"),
    )
    if periodo_tipo == "all":
        messages.error(request, "Selecciona periodo mensual o quincenal para guardar objetivo por categoría.")
        return redirect("compras:solicitudes")

    categoria = " ".join((request.POST.get("categoria") or "").strip().split())
    if not categoria:
        messages.error(request, "Ingresa una categoría válida.")
        return redirect("compras:solicitudes")

    monto_objetivo = _to_decimal(request.POST.get("monto_objetivo_categoria"), "0")
    if monto_objetivo < 0:
        monto_objetivo = Decimal("0")
    notas = (request.POST.get("notas_categoria") or "").strip()

    presupuesto_periodo, _ = PresupuestoCompraPeriodo.objects.get_or_create(
        periodo_tipo=periodo_tipo,
        periodo_mes=periodo_mes,
        defaults={"monto_objetivo": Decimal("0"), "actualizado_por": request.user},
    )
    objetivo_categoria, created = PresupuestoCompraCategoria.objects.update_or_create(
        presupuesto_periodo=presupuesto_periodo,
        categoria_normalizada=_normalize_categoria_text(categoria),
        defaults={
            "categoria": categoria,
            "monto_objetivo": monto_objetivo,
            "notas": notas,
            "actualizado_por": request.user,
        },
    )
    log_event(
        request.user,
        "CREATE" if created else "UPDATE",
        "compras.PresupuestoCompraCategoria",
        objetivo_categoria.id,
        {
            "periodo_tipo": periodo_tipo,
            "periodo_mes": periodo_mes,
            "categoria": categoria,
            "monto_objetivo": str(monto_objetivo),
        },
    )
    messages.success(request, f"Objetivo de categoría actualizado: {categoria}.")

    params = {
        "source": (request.POST.get("source") or "all").strip() or "all",
        "plan_id": (request.POST.get("plan_id") or "").strip(),
        "categoria": _sanitize_categoria_filter(request.POST.get("categoria")),
        "reabasto": (request.POST.get("reabasto") or "all").strip() or "all",
        "consumo_ref": _sanitize_consumo_ref_filter(request.POST.get("consumo_ref")),
        "periodo_tipo": periodo_tipo,
        "periodo_mes": periodo_mes,
    }
    if not params["plan_id"]:
        params.pop("plan_id")
    if not params["categoria"]:
        params.pop("categoria")
    return redirect(f"{reverse('compras:solicitudes')}?{urlencode(params)}")


@login_required
@require_POST
def importar_presupuestos_periodo(request: HttpRequest) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para importar presupuesto.")

    archivo = request.FILES.get("archivo_presupuesto")
    if not archivo:
        messages.error(request, "Selecciona un archivo de presupuesto.")
        return redirect("compras:solicitudes")

    try:
        rows = _read_import_rows(archivo)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("compras:solicitudes")
    except Exception:
        messages.error(request, "No se pudo leer el archivo de presupuesto.")
        return redirect("compras:solicitudes")

    if not rows:
        messages.warning(request, "El archivo de presupuesto no tiene filas.")
        return redirect("compras:solicitudes")

    providers_by_norm = {
        normalizar_nombre(p.nombre): p
        for p in Proveedor.objects.filter(activo=True).only("id", "nombre")
    }

    created = 0
    updated = 0
    created_proveedor = 0
    updated_proveedor = 0
    created_categoria = 0
    updated_categoria = 0
    skipped = 0
    for idx, row in enumerate(rows, start=2):
        periodo_tipo = _parse_periodo_tipo_value(row.get("periodo_tipo"))
        periodo_mes = _parse_periodo_mes_value(row.get("periodo_mes"))
        monto_raw = row.get("monto_objetivo")
        monto_has_value = str(monto_raw).strip() != "" if monto_raw is not None else False
        monto = _to_decimal(str(monto_raw or "0"), "0")
        notas = str(row.get("notas") or "").strip()

        if not periodo_tipo or not periodo_mes:
            skipped += 1
            continue
        if monto_has_value and monto < 0:
            monto = Decimal("0")

        presupuesto, was_created = PresupuestoCompraPeriodo.objects.get_or_create(
            periodo_tipo=periodo_tipo,
            periodo_mes=periodo_mes,
            defaults={
                "monto_objetivo": monto,
                "notas": notas,
                "actualizado_por": request.user,
            },
        )
        if monto_has_value:
            presupuesto.monto_objetivo = monto
            presupuesto.notas = notas
            presupuesto.actualizado_por = request.user
            presupuesto.save(update_fields=["monto_objetivo", "notas", "actualizado_por", "actualizado_en"])
            log_event(
                request.user,
                "CREATE" if was_created else "UPDATE",
                "compras.PresupuestoCompraPeriodo",
                presupuesto.id,
                {
                    "source": "import",
                    "row": idx,
                    "periodo_tipo": periodo_tipo,
                    "periodo_mes": periodo_mes,
                    "monto_objetivo": str(monto),
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

        had_dimension_update = False

        proveedor_raw = str(row.get("proveedor") or "").strip()
        if proveedor_raw:
            proveedor = providers_by_norm.get(normalizar_nombre(proveedor_raw))
            if not proveedor:
                skipped += 1
            else:
                monto_proveedor_raw = row.get("monto_objetivo_proveedor")
                monto_proveedor_has_value = (
                    str(monto_proveedor_raw).strip() != "" if monto_proveedor_raw is not None else False
                )
                if not monto_proveedor_has_value:
                    monto_proveedor_raw = monto_raw
                    monto_proveedor_has_value = monto_has_value

                if not monto_proveedor_has_value:
                    skipped += 1
                else:
                    monto_proveedor = _to_decimal(str(monto_proveedor_raw or "0"), "0")
                    if monto_proveedor < 0:
                        monto_proveedor = Decimal("0")
                    objetivo_proveedor, was_created_proveedor = PresupuestoCompraProveedor.objects.update_or_create(
                        presupuesto_periodo=presupuesto,
                        proveedor=proveedor,
                        defaults={
                            "monto_objetivo": monto_proveedor,
                            "notas": notas,
                            "actualizado_por": request.user,
                        },
                    )
                    log_event(
                        request.user,
                        "CREATE" if was_created_proveedor else "UPDATE",
                        "compras.PresupuestoCompraProveedor",
                        objetivo_proveedor.id,
                        {
                            "source": "import",
                            "row": idx,
                            "periodo_tipo": periodo_tipo,
                            "periodo_mes": periodo_mes,
                            "proveedor_id": proveedor.id,
                            "proveedor_nombre": proveedor.nombre,
                            "monto_objetivo": str(monto_proveedor),
                        },
                    )
                    had_dimension_update = True
                    if was_created_proveedor:
                        created_proveedor += 1
                    else:
                        updated_proveedor += 1

        categoria_raw = " ".join((str(row.get("categoria") or "")).strip().split())
        if categoria_raw:
            categoria_norm = _normalize_categoria_text(categoria_raw)
            monto_categoria_raw = row.get("monto_objetivo_categoria")
            monto_categoria_has_value = (
                str(monto_categoria_raw).strip() != "" if monto_categoria_raw is not None else False
            )
            if not monto_categoria_has_value:
                monto_categoria_raw = monto_raw
                monto_categoria_has_value = monto_has_value

            if not monto_categoria_has_value:
                skipped += 1
            else:
                monto_categoria = _to_decimal(str(monto_categoria_raw or "0"), "0")
                if monto_categoria < 0:
                    monto_categoria = Decimal("0")
                objetivo_categoria, was_created_categoria = PresupuestoCompraCategoria.objects.update_or_create(
                    presupuesto_periodo=presupuesto,
                    categoria_normalizada=categoria_norm,
                    defaults={
                        "categoria": categoria_raw,
                        "monto_objetivo": monto_categoria,
                        "notas": notas,
                        "actualizado_por": request.user,
                    },
                )
                log_event(
                    request.user,
                    "CREATE" if was_created_categoria else "UPDATE",
                    "compras.PresupuestoCompraCategoria",
                    objetivo_categoria.id,
                    {
                        "source": "import",
                        "row": idx,
                        "periodo_tipo": periodo_tipo,
                        "periodo_mes": periodo_mes,
                        "categoria": categoria_raw,
                        "monto_objetivo": str(monto_categoria),
                    },
                )
                had_dimension_update = True
                if was_created_categoria:
                    created_categoria += 1
                else:
                    updated_categoria += 1

        if (not monto_has_value) and (not had_dimension_update):
            skipped += 1

    messages.success(
        request,
        (
            "Importación de presupuesto completada. "
            f"Periodo nuevos: {created}. "
            f"Periodo actualizados: {updated}. "
            f"Proveedor nuevos: {created_proveedor}. "
            f"Proveedor actualizados: {updated_proveedor}. "
            f"Categoría nuevas: {created_categoria}. "
            f"Categoría actualizadas: {updated_categoria}. "
            f"Omitidos: {skipped}."
        ),
    )

    params = {
        "source": (request.POST.get("source") or "all").strip() or "all",
        "plan_id": (request.POST.get("plan_id") or "").strip(),
        "categoria": _sanitize_categoria_filter(request.POST.get("categoria")),
        "reabasto": (request.POST.get("reabasto") or "all").strip() or "all",
        "periodo_tipo": (request.POST.get("periodo_tipo") or "mes").strip() or "mes",
        "periodo_mes": (request.POST.get("periodo_mes") or "").strip(),
    }
    if not params["plan_id"]:
        params.pop("plan_id")
    if not params["categoria"]:
        params.pop("categoria")
    if not params["periodo_mes"]:
        params.pop("periodo_mes")
    return redirect(f"{reverse('compras:solicitudes')}?{urlencode(params)}")


@login_required
@require_POST
def importar_solicitudes(request: HttpRequest) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para importar solicitudes.")

    archivo = request.FILES.get("archivo")
    if not archivo:
        messages.error(request, "Debes seleccionar un archivo de importación (.xlsx o .csv).")
        return redirect("compras:solicitudes")

    periodo_tipo, periodo_mes, _ = _parse_period_filters(
        request.POST.get("periodo_tipo"),
        request.POST.get("periodo_mes"),
    )
    fecha_default = _default_fecha_requerida(periodo_tipo, periodo_mes)
    area_default = (request.POST.get("area") or "General").strip() or "General"
    solicitante_default = (request.POST.get("solicitante") or request.user.username).strip() or request.user.username
    estatus_default = (request.POST.get("estatus") or SolicitudCompra.STATUS_BORRADOR).strip().upper()
    valid_status = {x[0] for x in SolicitudCompra.STATUS_CHOICES}
    if estatus_default not in valid_status:
        estatus_default = SolicitudCompra.STATUS_BORRADOR
    evitar_duplicados = request.POST.get("evitar_duplicados") == "on"
    min_score_raw = request.POST.get("score_min") or "90"
    try:
        min_score = max(0, min(100, int(min_score_raw)))
    except ValueError:
        min_score = 90

    try:
        rows = _read_import_rows(archivo)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("compras:solicitudes")
    except Exception:
        messages.error(request, "No se pudo leer el archivo. Verifica formato y columnas.")
        return redirect("compras:solicitudes")

    if not rows:
        messages.warning(request, "El archivo no contiene filas de datos.")
        return redirect("compras:solicitudes")

    provider_map = {
        normalizar_nombre(p.nombre): p
        for p in Proveedor.objects.filter(activo=True).only("id", "nombre")
    }
    match_cache: dict[str, tuple[Insumo | None, float, str]] = {}
    parsed_rows: list[dict] = []
    duplicate_keys_to_check: set[tuple[str, int, date]] = set()

    for idx, row in enumerate(rows, start=2):
        insumo_raw = str(row.get("insumo") or "").strip()
        cantidad_raw = str(row.get("cantidad") or "").strip()
        cantidad = _to_decimal(cantidad_raw, "0")
        if insumo_raw:
            cache_key = normalizar_nombre(insumo_raw)
            insumo_match, score, method = match_cache.get(cache_key, (None, 0.0, "sin_match"))
            if cache_key not in match_cache:
                insumo_match, score, method = match_insumo(insumo_raw)
                insumo_match = canonical_insumo(insumo_match)
                match_cache[cache_key] = (insumo_match, score, method)
        else:
            insumo_match, score, method = (None, 0.0, "sin_match")
        insumo_id = int(insumo_match.id) if (insumo_match and score >= min_score) else 0

        area = str(row.get("area") or area_default).strip() or area_default
        solicitante = str(row.get("solicitante") or solicitante_default).strip() or solicitante_default
        fecha_requerida = _parse_date_value(row.get("fecha_requerida"), fecha_default)
        estatus = str(row.get("estatus") or estatus_default).strip().upper()
        if estatus not in valid_status:
            estatus = estatus_default
        proveedor = _resolve_proveedor_name(str(row.get("proveedor") or ""), provider_map)
        if not proveedor and insumo_match:
            proveedor = insumo_match.proveedor_principal

        parsed_rows.append(
            {
                "row_id": str(idx),
                "source_row": idx,
                "insumo_origen": insumo_raw,
                "insumo_sugerencia": insumo_match.nombre if insumo_match else "",
                "insumo_id": insumo_id,
                "cantidad": cantidad,
                "cantidad_origen": cantidad_raw,
                "area": area,
                "solicitante": solicitante,
                "fecha_requerida": fecha_requerida,
                "estatus": estatus,
                "proveedor_id": int(proveedor.id) if proveedor else 0,
                "score": float(score or 0),
                "metodo": method,
                "has_insumo_match": bool(insumo_match),
            }
        )
        if evitar_duplicados and insumo_id:
            duplicate_keys_to_check.add((area, insumo_id, fecha_requerida))

    duplicates_found: set[tuple[str, int, date]] = set()
    if duplicate_keys_to_check:
        areas = sorted({k[0] for k in duplicate_keys_to_check})
        insumo_ids = sorted({k[1] for k in duplicate_keys_to_check})
        fechas = sorted({k[2] for k in duplicate_keys_to_check})
        duplicates_found = {
            (area, int(insumo_id), fecha)
            for area, insumo_id, fecha in SolicitudCompra.objects.filter(
                area__in=areas,
                insumo_id__in=insumo_ids,
                fecha_requerida__in=fechas,
                estatus__in=_active_solicitud_statuses(),
            ).values_list("area", "insumo_id", "fecha_requerida")
        }

    preview_cost_by_insumo: dict[int, Decimal] = {}
    preview_insumo_ids = sorted({int(p["insumo_id"]) for p in parsed_rows if int(p["insumo_id"] or 0) > 0})
    for insumo_id in preview_insumo_ids:
        canonical = canonical_insumo_by_id(insumo_id)
        latest = latest_costo_canonico(insumo_id=canonical.id if canonical else insumo_id)
        if latest is not None:
            preview_cost_by_insumo[insumo_id] = latest

    preview_rows: list[dict] = []
    for parsed in parsed_rows:
        insumo_raw = str(parsed["insumo_origen"] or "").strip()
        cantidad = parsed["cantidad"]
        cantidad_raw = parsed["cantidad_origen"]
        insumo_id = parsed["insumo_id"]
        costo_unitario = preview_cost_by_insumo.get(insumo_id, Decimal("0")) if insumo_id else Decimal("0")
        presupuesto_estimado = (cantidad * costo_unitario) if cantidad > 0 else Decimal("0")
        area = parsed["area"]
        solicitante = parsed["solicitante"]
        fecha_requerida = parsed["fecha_requerida"]
        estatus = parsed["estatus"]
        duplicate = bool(insumo_id and (area, insumo_id, fecha_requerida) in duplicates_found)

        notes: list[str] = []
        hard_error = False
        if not insumo_raw:
            notes.append("Insumo vacío en archivo.")
            hard_error = True
        if not insumo_id:
            if parsed["has_insumo_match"]:
                notes.append(f"Artículo detectado con confianza menor a {min_score}.")
            else:
                notes.append("Artículo no identificado en el catálogo.")
            hard_error = True
        if cantidad <= 0:
            notes.append("Cantidad inválida (debe ser > 0).")
            hard_error = True
        if duplicate:
            notes.append("Posible duplicado con solicitud activa.")

        preview_rows.append(
            {
                "row_id": parsed["row_id"],
                "source_row": parsed["source_row"],
                "insumo_origen": insumo_raw,
                "insumo_sugerencia": parsed["insumo_sugerencia"],
                "insumo_id": str(insumo_id) if insumo_id else "",
                "cantidad": str(cantidad),
                "cantidad_origen": cantidad_raw,
                "area": area,
                "solicitante": solicitante,
                "fecha_requerida": fecha_requerida.isoformat(),
                "estatus": estatus,
                "proveedor_id": str(parsed["proveedor_id"]) if parsed["proveedor_id"] else "",
                "score": f"{parsed['score']:.1f}",
                "metodo": parsed["metodo"],
                "costo_unitario": str(costo_unitario),
                "presupuesto_estimado": str(presupuesto_estimado),
                "duplicate": duplicate,
                "notes": " | ".join(notes),
                "include": not hard_error,
            }
        )

    request.session[IMPORT_PREVIEW_SESSION_KEY] = {
        "periodo_tipo": periodo_tipo,
        "periodo_mes": periodo_mes,
        "evitar_duplicados": evitar_duplicados,
        "score_min": min_score,
        "file_name": archivo.name,
        "generated_at": timezone.localtime().strftime("%Y-%m-%d %H:%M"),
        "rows": preview_rows,
    }
    request.session.modified = True
    messages.info(
        request,
        f"Vista previa generada con {len(preview_rows)} filas. Edita, elimina filas y confirma la importación.",
    )

    return redirect(
        f"{reverse('compras:solicitudes')}?source=manual&periodo_tipo={periodo_tipo}&periodo_mes={periodo_mes}"
    )


@login_required
@require_POST
def confirmar_importacion_solicitudes(request: HttpRequest) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para importar solicitudes.")

    preview_payload = request.session.get(IMPORT_PREVIEW_SESSION_KEY)
    if not preview_payload:
        messages.error(request, "No hay una vista previa activa para confirmar.")
        return redirect("compras:solicitudes")

    rows = preview_payload.get("rows") or []
    periodo_tipo = (preview_payload.get("periodo_tipo") or "mes").strip() or "mes"
    periodo_mes = (preview_payload.get("periodo_mes") or "").strip()
    evitar_duplicados = bool(preview_payload.get("evitar_duplicados"))
    default_fecha = _default_fecha_requerida(periodo_tipo, periodo_mes)
    valid_status = {x[0] for x in SolicitudCompra.STATUS_CHOICES}

    raw_insumo_ids: set[int] = set()
    proveedor_ids: set[int] = set()
    for row in rows:
        row_id = str(row.get("row_id") or "")
        if request.POST.get(f"row_{row_id}_include") != "on":
            continue
        try:
            raw_insumo_ids.add(int(request.POST.get(f"row_{row_id}_insumo_id") or "0"))
        except ValueError:
            pass
        try:
            proveedor_ids.add(int(request.POST.get(f"row_{row_id}_proveedor_id") or "0"))
        except ValueError:
            pass

    canonical_by_posted_id = {
        insumo_id: canonical_insumo_by_id(insumo_id)
        for insumo_id in raw_insumo_ids
        if insumo_id > 0
    }
    canonical_ids = sorted({insumo.id for insumo in canonical_by_posted_id.values() if insumo})
    insumos_map = Insumo.objects.select_related("proveedor_principal").in_bulk(canonical_ids)
    proveedores_map = Proveedor.objects.filter(activo=True, id__in=proveedor_ids).in_bulk()

    existing_duplicate_keys: set[tuple[str, int, date]] = set()
    if evitar_duplicados:
        batch_keys: set[tuple[str, int, date]] = set()
        for row in rows:
            row_id = str(row.get("row_id") or "")
            if request.POST.get(f"row_{row_id}_include") != "on":
                continue
            area = (request.POST.get(f"row_{row_id}_area") or "").strip() or "General"
            try:
                posted_insumo_id = int(request.POST.get(f"row_{row_id}_insumo_id") or "0")
            except ValueError:
                posted_insumo_id = 0
            insumo_id = canonical_by_posted_id.get(posted_insumo_id).id if canonical_by_posted_id.get(posted_insumo_id) else 0
            if insumo_id <= 0:
                continue
            fecha_requerida = _parse_date_value(request.POST.get(f"row_{row_id}_fecha_requerida"), default_fecha)
            batch_keys.add((area, insumo_id, fecha_requerida))
        if batch_keys:
            areas = sorted({k[0] for k in batch_keys})
            insumo_ids_batch = sorted({k[1] for k in batch_keys})
            fechas = sorted({k[2] for k in batch_keys})
            existing_duplicate_keys = {
                (area, int(insumo_id), fecha)
                for area, insumo_id, fecha in SolicitudCompra.objects.filter(
                    area__in=areas,
                    insumo_id__in=insumo_ids_batch,
                    fecha_requerida__in=fechas,
                    estatus__in=_active_solicitud_statuses(),
                ).values_list("area", "insumo_id", "fecha_requerida")
            }

    created = 0
    skipped_invalid = 0
    skipped_duplicate = 0
    skipped_removed = 0
    created_duplicate_keys: set[tuple[str, int, date]] = set()

    for row in rows:
        row_id = str(row.get("row_id") or "")
        if request.POST.get(f"row_{row_id}_include") != "on":
            skipped_removed += 1
            continue

        area = (request.POST.get(f"row_{row_id}_area") or "").strip() or "General"
        solicitante = (request.POST.get(f"row_{row_id}_solicitante") or "").strip() or request.user.username
        estatus = (request.POST.get(f"row_{row_id}_estatus") or SolicitudCompra.STATUS_BORRADOR).strip().upper()
        if estatus not in valid_status:
            estatus = SolicitudCompra.STATUS_BORRADOR

        try:
            posted_insumo_id = int(request.POST.get(f"row_{row_id}_insumo_id") or "0")
        except ValueError:
            posted_insumo_id = 0
        canonical_insumo_obj = canonical_by_posted_id.get(posted_insumo_id)
        insumo = insumos_map.get(canonical_insumo_obj.id) if canonical_insumo_obj else None
        if not insumo:
            skipped_invalid += 1
            continue

        cantidad = _to_decimal(request.POST.get(f"row_{row_id}_cantidad"), "0")
        if cantidad <= 0:
            skipped_invalid += 1
            continue

        fecha_requerida = _parse_date_value(request.POST.get(f"row_{row_id}_fecha_requerida"), default_fecha)

        proveedor = None
        try:
            proveedor_id = int(request.POST.get(f"row_{row_id}_proveedor_id") or "0")
            proveedor = proveedores_map.get(proveedor_id)
        except ValueError:
            proveedor = None
        if not proveedor:
            proveedor = insumo.proveedor_principal

        duplicate_key = (area, int(insumo.id), fecha_requerida)
        if evitar_duplicados:
            if (duplicate_key in existing_duplicate_keys) or (duplicate_key in created_duplicate_keys):
                skipped_duplicate += 1
                continue

        solicitud = SolicitudCompra.objects.create(
            area=area,
            solicitante=solicitante,
            insumo=insumo,
            proveedor_sugerido=proveedor,
            cantidad=cantidad,
            fecha_requerida=fecha_requerida,
            estatus=estatus,
        )
        log_event(
            request.user,
            "CREATE",
            "compras.SolicitudCompra",
            solicitud.id,
            {
                "folio": solicitud.folio,
                "source": "import_preview_confirm",
            },
        )
        if evitar_duplicados:
            created_duplicate_keys.add(duplicate_key)
        created += 1

    request.session.pop(IMPORT_PREVIEW_SESSION_KEY, None)
    request.session.modified = True
    messages.success(
        request,
        (
            f"Importación confirmada. Creadas: {created}. "
            f"Eliminadas en vista previa: {skipped_removed}. "
            f"Duplicadas: {skipped_duplicate}. Inválidas: {skipped_invalid}."
        ),
    )
    return redirect(
        f"{reverse('compras:solicitudes')}?source=manual&periodo_tipo={periodo_tipo}&periodo_mes={periodo_mes}"
    )


@login_required
@require_POST
def cancelar_importacion_solicitudes(request: HttpRequest) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para importar solicitudes.")
    request.session.pop(IMPORT_PREVIEW_SESSION_KEY, None)
    request.session.modified = True
    messages.info(request, "Vista previa de importación eliminada.")
    return redirect("compras:solicitudes")


@login_required
def solicitudes_print(request: HttpRequest) -> HttpResponse:
    if not can_view_compras(request.user):
        raise PermissionDenied("No tienes permisos para ver Compras.")

    (
        solicitudes,
        source_filter,
        plan_filter,
        categoria_filter,
        reabasto_filter,
        estatus_filter,
        _,
        _,
        _,
        periodo_tipo,
        periodo_mes,
        periodo_label,
    ) = _filtered_solicitudes(
        request.GET.get("source"),
        request.GET.get("plan_id"),
        request.GET.get("categoria"),
        request.GET.get("reabasto"),
        request.GET.get("estatus"),
        request.GET.get("workflow_action"),
        request.GET.get("blocker_key"),
        request.GET.get("periodo_tipo"),
        request.GET.get("periodo_mes"),
    )

    total_cantidad = sum((s.cantidad for s in solicitudes), Decimal("0"))
    total_presupuesto = sum((s.presupuesto_estimado for s in solicitudes), Decimal("0"))
    criticos_count = sum(1 for s in solicitudes if s.reabasto_nivel == "critico")
    bajos_count = sum(1 for s in solicitudes if s.reabasto_nivel == "bajo")
    ok_count = sum(1 for s in solicitudes if s.reabasto_nivel == "ok")

    context = {
        "solicitudes": solicitudes,
        "source_filter": source_filter,
        "plan_filter": plan_filter or "-",
        "categoria_filter": categoria_filter or "-",
        "reabasto_filter": reabasto_filter,
        "periodo_label": periodo_label,
        "periodo_mes": periodo_mes if periodo_tipo != "all" else "-",
        "total_cantidad": total_cantidad,
        "total_presupuesto": total_presupuesto,
        "criticos_count": criticos_count,
        "bajos_count": bajos_count,
        "ok_count": ok_count,
        "generated_at": timezone.localtime(),
        "generated_by": request.user.username,
        "document_folio": _solicitudes_print_folio(),
        "status_autorizacion": "Pendiente de firmas",
        "return_query": request.GET.urlencode(),
    }
    return render(request, "compras/solicitudes_print.html", context)


@login_required
def solicitudes_resumen_api(request: HttpRequest) -> JsonResponse:
    if not can_view_compras(request.user):
        raise PermissionDenied("No tienes permisos para ver Compras.")

    (
        solicitudes,
        source_filter,
        plan_filter,
        categoria_filter,
        reabasto_filter,
        estatus_filter,
        _,
        _,
        _,
        periodo_tipo,
        periodo_mes,
        periodo_label,
    ) = _filtered_solicitudes(
        request.GET.get("source"),
        request.GET.get("plan_id"),
        request.GET.get("categoria"),
        request.GET.get("reabasto"),
        request.GET.get("estatus"),
        request.GET.get("workflow_action"),
        request.GET.get("blocker_key"),
        request.GET.get("periodo_tipo"),
        request.GET.get("periodo_mes"),
    )
    consumo_ref_filter = _sanitize_consumo_ref_filter(request.GET.get("consumo_ref"))

    budget_ctx = _build_budget_context(
        solicitudes,
        source_filter,
        plan_filter,
        categoria_filter,
        periodo_tipo,
        periodo_mes,
    )
    provider_dashboard = _build_provider_dashboard(
        periodo_mes,
        source_filter,
        plan_filter,
        categoria_filter,
        budget_ctx["presupuesto_rows_proveedor"],
    )
    category_dashboard = _build_category_dashboard(
        periodo_mes,
        source_filter,
        plan_filter,
        categoria_filter,
        budget_ctx.get("presupuesto_rows_categoria", []),
    )
    consumo_dashboard = _build_consumo_vs_plan_dashboard(
        periodo_tipo,
        periodo_mes,
        source_filter,
        plan_filter,
        categoria_filter,
        consumo_ref_filter,
    )
    historial = _build_budget_history(periodo_mes, source_filter, plan_filter, categoria_filter)

    data = {
        "filters": {
            "source": source_filter,
            "plan_id": plan_filter or "",
            "categoria": categoria_filter or "",
            "reabasto": reabasto_filter,
            "consumo_ref": consumo_ref_filter,
            "periodo_tipo": periodo_tipo,
            "periodo_mes": periodo_mes,
            "periodo_label": periodo_label,
        },
        "totals": {
            "solicitudes_count": len(solicitudes),
            "presupuesto_estimado_total": float(budget_ctx["presupuesto_estimado_total"] or 0),
            "presupuesto_ejecutado_total": float(budget_ctx["presupuesto_ejecutado_total"] or 0),
            "presupuesto_objetivo": float((budget_ctx.get("presupuesto_objetivo") or 0)),
            "presupuesto_variacion_objetivo": float((budget_ctx.get("presupuesto_variacion_objetivo") or 0)),
            "alertas_total": int(budget_ctx.get("presupuesto_alertas_total") or 0),
            "alertas_excedidas": int(budget_ctx.get("presupuesto_alertas_excedidas") or 0),
            "alertas_preventivas": int(budget_ctx.get("presupuesto_alertas_preventivas") or 0),
            "proveedor_objetivo_excedido_count": int(budget_ctx.get("presupuesto_proveedor_excedido_count") or 0),
            "proveedor_objetivo_preventivo_count": int(budget_ctx.get("presupuesto_proveedor_preventivo_count") or 0),
            "categoria_objetivo_excedido_count": int(budget_ctx.get("presupuesto_categoria_excedido_count") or 0),
            "categoria_objetivo_preventivo_count": int(budget_ctx.get("presupuesto_categoria_preventivo_count") or 0),
        },
        "top_proveedores": [
            {
                "proveedor": row["proveedor"],
                "estimado": float(row["estimado"] or 0),
                "ejecutado": float(row["ejecutado"] or 0),
                "variacion": float(row["variacion"] or 0),
                "participacion_pct": float(row["participacion_pct"] or 0),
                "objetivo_proveedor": float((row.get("objetivo_proveedor") or 0)),
                "uso_objetivo_pct": (
                    float(row["uso_objetivo_pct"])
                    if row.get("uso_objetivo_pct") is not None
                    else None
                ),
                "objetivo_estado": row.get("objetivo_estado") or "",
                "objetivo_estado_label": row.get("objetivo_estado_label") or "",
            }
            for row in budget_ctx["presupuesto_rows_proveedor"][:10]
        ],
        "top_categorias": [
            {
                "categoria": row["categoria"],
                "estimado": float(row["estimado"] or 0),
                "ejecutado": float(row["ejecutado"] or 0),
                "variacion": float(row["variacion"] or 0),
                "participacion_pct": float(row["participacion_pct"] or 0),
                "objetivo_categoria": float((row.get("objetivo_categoria") or 0)),
                "uso_objetivo_pct": (
                    float(row["uso_objetivo_pct"])
                    if row.get("uso_objetivo_pct") is not None
                    else None
                ),
                "objetivo_estado": row.get("objetivo_estado") or "",
                "objetivo_estado_label": row.get("objetivo_estado_label") or "",
            }
            for row in budget_ctx.get("presupuesto_rows_categoria", [])[:10]
        ],
        "historial_6m": [
            {
                "periodo_mes": row["periodo_mes"],
                "objetivo": float(row["objetivo"] or 0),
                "estimado": float(row["estimado"] or 0),
                "ejecutado": float(row["ejecutado"] or 0),
                "ratio_pct": float(row["ratio_pct"]) if row.get("ratio_pct") is not None else None,
                "estado_label": row["estado_label"],
            }
            for row in historial
        ],
        "trend": {
            "proveedor_rows": [
                {
                    "proveedor": row["proveedor"],
                    "mes": row["mes"],
                    "estimado": float(row["estimado"] or 0),
                    "ejecutado": float(row["ejecutado"] or 0),
                    "variacion": float(row["variacion"] or 0),
                }
                for row in provider_dashboard["trend_rows"]
            ],
            "categoria_rows": [
                {
                    "categoria": row["categoria"],
                    "mes": row["mes"],
                    "estimado": float(row["estimado"] or 0),
                    "ejecutado": float(row["ejecutado"] or 0),
                    "variacion": float(row["variacion"] or 0),
                }
                for row in category_dashboard["trend_rows"]
            ],
        },
        "consumo_vs_plan": {
            "totals": {
                "plan_qty_total": float(consumo_dashboard["totals"]["plan_qty_total"] or 0),
                "consumo_real_qty_total": float(consumo_dashboard["totals"]["consumo_real_qty_total"] or 0),
                "plan_cost_total": float(consumo_dashboard["totals"]["plan_cost_total"] or 0),
                "consumo_real_cost_total": float(consumo_dashboard["totals"]["consumo_real_cost_total"] or 0),
                "variacion_cost_total": float(consumo_dashboard["totals"]["variacion_cost_total"] or 0),
                "sin_costo_count": int(consumo_dashboard["totals"]["sin_costo_count"] or 0),
                "semaforo_verde_count": int(consumo_dashboard["totals"]["semaforo_verde_count"] or 0),
                "semaforo_amarillo_count": int(consumo_dashboard["totals"]["semaforo_amarillo_count"] or 0),
                "semaforo_rojo_count": int(consumo_dashboard["totals"]["semaforo_rojo_count"] or 0),
                "cobertura_pct": (
                    float(consumo_dashboard["totals"]["cobertura_pct"])
                    if consumo_dashboard["totals"]["cobertura_pct"] is not None
                    else None
                ),
            },
            "period_start": str(consumo_dashboard["period_start"]),
            "period_end": str(consumo_dashboard["period_end"]),
            "rows": [
                {
                    "insumo_id": row["insumo_id"],
                    "insumo": row["insumo"],
                    "categoria": row["categoria"],
                    "unidad": row["unidad"],
                    "cantidad_plan": float(row["cantidad_plan"] or 0),
                    "cantidad_real": float(row["cantidad_real"] or 0),
                    "variacion_qty": float(row["variacion_qty"] or 0),
                    "costo_unitario": float(row["costo_unitario"] or 0),
                    "costo_plan": float(row["costo_plan"] or 0),
                    "costo_real": float(row["costo_real"] or 0),
                    "variacion_cost": float(row["variacion_cost"] or 0),
                    "consumo_pct": float(row["consumo_pct"]) if row["consumo_pct"] is not None else None,
                    "estado": row["estado"],
                    "semaforo": row["semaforo"],
                    "sin_costo": bool(row["sin_costo"]),
                    "alerta": row["alerta"],
                }
                for row in consumo_dashboard["rows"][:20]
            ],
        },
    }
    return JsonResponse(data)


@login_required
def solicitudes_consumo_vs_plan_api(request: HttpRequest) -> JsonResponse:
    if not can_view_compras(request.user):
        raise PermissionDenied("No tienes permisos para ver Compras.")

    periodo_tipo, periodo_mes, periodo_label = _parse_period_filters(
        request.GET.get("periodo_tipo"),
        request.GET.get("periodo_mes"),
    )
    source_filter = (request.GET.get("source") or "all").lower()
    if source_filter not in {"all", "manual", "plan"}:
        source_filter = "all"
    plan_filter = (request.GET.get("plan_id") or "").strip()
    categoria_filter = _sanitize_categoria_filter(request.GET.get("categoria"))
    consumo_ref_filter = _sanitize_consumo_ref_filter(request.GET.get("consumo_ref"))
    try:
        limit = int(request.GET.get("limit", 30))
    except (TypeError, ValueError):
        limit = 30
    limit = max(1, min(limit, 1000))
    try:
        offset = int(request.GET.get("offset", 0))
    except (TypeError, ValueError):
        offset = 0
    offset = max(0, min(offset, 200000))
    sort_by = (request.GET.get("sort_by") or "variacion_cost_abs").strip().lower()
    sort_dir = (request.GET.get("sort_dir") or "desc").strip().lower()
    allowed_sort = {
        "variacion_cost_abs",
        "variacion_cost",
        "costo_real",
        "costo_plan",
        "cantidad_real",
        "cantidad_plan",
        "consumo_pct",
        "insumo",
        "categoria",
        "estado",
        "semaforo",
    }
    if sort_by not in allowed_sort:
        return JsonResponse(
            {"detail": "sort_by inválido."},
            status=400,
        )
    if sort_dir not in {"asc", "desc"}:
        return JsonResponse(
            {"detail": "sort_dir inválido."},
            status=400,
        )

    dashboard = _build_consumo_vs_plan_dashboard(
        periodo_tipo=periodo_tipo,
        periodo_mes=periodo_mes,
        source_filter=source_filter,
        plan_filter=plan_filter,
        categoria_filter=categoria_filter,
        consumo_ref_filter=consumo_ref_filter,
        limit=limit,
        offset=offset,
        sort_by=sort_by,
        sort_dir=sort_dir,
    )
    data = {
        "filters": {
            "source": source_filter,
            "plan_id": plan_filter,
            "categoria": categoria_filter or "",
            "consumo_ref": consumo_ref_filter,
            "periodo_tipo": periodo_tipo,
            "periodo_mes": periodo_mes,
            "periodo_label": periodo_label,
            "limit": limit,
            "offset": offset,
            "sort_by": sort_by,
            "sort_dir": sort_dir,
        },
        "period_start": str(dashboard["period_start"]),
        "period_end": str(dashboard["period_end"]),
        "meta": {
            "rows_total": int(dashboard.get("meta", {}).get("rows_total") or 0),
            "rows_returned": int(dashboard.get("meta", {}).get("rows_returned") or 0),
            "limit": int(dashboard.get("meta", {}).get("limit") or limit),
            "offset": int(dashboard.get("meta", {}).get("offset") or offset),
            "sort_by": dashboard.get("meta", {}).get("sort_by") or sort_by,
            "sort_dir": dashboard.get("meta", {}).get("sort_dir") or sort_dir,
        },
        "totals": {
            "plan_qty_total": float(dashboard["totals"]["plan_qty_total"] or 0),
            "consumo_real_qty_total": float(dashboard["totals"]["consumo_real_qty_total"] or 0),
            "plan_cost_total": float(dashboard["totals"]["plan_cost_total"] or 0),
            "consumo_real_cost_total": float(dashboard["totals"]["consumo_real_cost_total"] or 0),
            "variacion_cost_total": float(dashboard["totals"]["variacion_cost_total"] or 0),
            "sin_costo_count": int(dashboard["totals"]["sin_costo_count"] or 0),
            "semaforo_verde_count": int(dashboard["totals"]["semaforo_verde_count"] or 0),
            "semaforo_amarillo_count": int(dashboard["totals"]["semaforo_amarillo_count"] or 0),
            "semaforo_rojo_count": int(dashboard["totals"]["semaforo_rojo_count"] or 0),
            "cobertura_pct": (
                float(dashboard["totals"]["cobertura_pct"])
                if dashboard["totals"]["cobertura_pct"] is not None
                else None
            ),
        },
        "rows": [
            {
                "insumo_id": row["insumo_id"],
                "insumo": row["insumo"],
                "categoria": row["categoria"],
                "unidad": row["unidad"],
                "cantidad_plan": float(row["cantidad_plan"] or 0),
                "cantidad_real": float(row["cantidad_real"] or 0),
                "variacion_qty": float(row["variacion_qty"] or 0),
                "costo_unitario": float(row["costo_unitario"] or 0),
                "costo_plan": float(row["costo_plan"] or 0),
                "costo_real": float(row["costo_real"] or 0),
                "variacion_cost": float(row["variacion_cost"] or 0),
                "consumo_pct": float(row["consumo_pct"]) if row["consumo_pct"] is not None else None,
                "estado": row["estado"],
                "semaforo": row["semaforo"],
                "sin_costo": bool(row["sin_costo"]),
                "alerta": row["alerta"],
            }
            for row in dashboard["rows"]
        ],
    }
    return JsonResponse(data)


def _parse_month_filter(raw: str) -> tuple[int, int] | None:
    value = (raw or "").strip()
    if not value:
        return None
    try:
        year_str, month_str = value.split("-")
        year = int(year_str)
        month = int(month_str)
        if 1 <= month <= 12:
            return year, month
    except Exception:
        return None
    return None


def _export_ordenes_csv(ordenes_qs) -> HttpResponse:
    ordenes = list(ordenes_qs)
    plan_ids = {
        int(plan_id)
        for plan_id in {
            _extract_plan_id_from_scope(orden.referencia)
            or _extract_plan_id_from_scope(getattr(orden.solicitud, "area", ""))
            for orden in ordenes
        }
        if plan_id
    }
    planes_map = {
        p.id: p
        for p in PlanProduccion.objects.filter(id__in=plan_ids).only("id", "nombre", "notas")
    }
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="ordenes_compra_{now_str}.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "folio",
            "solicitud_folio",
            "origen",
            "plan_origen",
            "referencia",
            "proveedor",
            "fecha_emision",
            "fecha_entrega_estimada",
            "monto_estimado",
            "estatus",
        ]
    )
    for orden in ordenes:
        source = _source_context_from_scope(
            area=getattr(orden.solicitud, "area", ""),
            referencia=orden.referencia,
            planes_map=planes_map,
        )
        writer.writerow(
            [
                orden.folio,
                orden.solicitud.folio if orden.solicitud else "",
                source["source_tipo"],
                source["source_plan_nombre"] or source["source_label"],
                orden.referencia or "",
                orden.proveedor.nombre,
                orden.fecha_emision.isoformat() if orden.fecha_emision else "",
                orden.fecha_entrega_estimada.isoformat() if orden.fecha_entrega_estimada else "",
                f"{Decimal(orden.monto_estimado or 0):.2f}",
                orden.get_estatus_display(),
            ]
        )
    return response


def _export_ordenes_xlsx(ordenes_qs) -> HttpResponse:
    ordenes = list(ordenes_qs)
    plan_ids = {
        int(plan_id)
        for plan_id in {
            _extract_plan_id_from_scope(orden.referencia)
            or _extract_plan_id_from_scope(getattr(orden.solicitud, "area", ""))
            for orden in ordenes
        }
        if plan_id
    }
    planes_map = {
        p.id: p
        for p in PlanProduccion.objects.filter(id__in=plan_ids).only("id", "nombre", "notas")
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "ordenes_compra"
    ws.append(
        [
            "folio",
            "solicitud_folio",
            "origen",
            "plan_origen",
            "referencia",
            "proveedor",
            "fecha_emision",
            "fecha_entrega_estimada",
            "monto_estimado",
            "estatus",
        ]
    )
    for orden in ordenes:
        source = _source_context_from_scope(
            area=getattr(orden.solicitud, "area", ""),
            referencia=orden.referencia,
            planes_map=planes_map,
        )
        ws.append(
            [
                orden.folio,
                orden.solicitud.folio if orden.solicitud else "",
                source["source_tipo"],
                source["source_plan_nombre"] or source["source_label"],
                orden.referencia or "",
                orden.proveedor.nombre,
                orden.fecha_emision.isoformat() if orden.fecha_emision else "",
                orden.fecha_entrega_estimada.isoformat() if orden.fecha_entrega_estimada else "",
                float(Decimal(orden.monto_estimado or 0)),
                orden.get_estatus_display(),
            ]
        )
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response["Content-Disposition"] = f'attachment; filename="ordenes_compra_{now_str}.xlsx"'
    return response


def _export_recepciones_csv(recepciones_qs) -> HttpResponse:
    recepciones = list(recepciones_qs)
    plan_ids = set()
    for rec in recepciones:
        solicitud_area = getattr(getattr(rec.orden, "solicitud", None), "area", "")
        plan_id = _extract_plan_id_from_scope(rec.orden.referencia) or _extract_plan_id_from_scope(solicitud_area)
        if plan_id:
            plan_ids.add(int(plan_id))
    planes_map = {
        p.id: p
        for p in PlanProduccion.objects.filter(id__in=plan_ids).only("id", "nombre", "notas")
    }
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="recepciones_compra_{now_str}.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "folio",
            "orden_folio",
            "origen",
            "plan_origen",
            "proveedor",
            "fecha_recepcion",
            "conformidad_pct",
            "estatus",
            "observaciones",
        ]
    )
    for rec in recepciones:
        source = _source_context_from_scope(
            area=getattr(getattr(rec.orden, "solicitud", None), "area", ""),
            referencia=rec.orden.referencia,
            planes_map=planes_map,
        )
        writer.writerow(
            [
                rec.folio,
                rec.orden.folio,
                source["source_tipo"],
                source["source_plan_nombre"] or source["source_label"],
                rec.orden.proveedor.nombre,
                rec.fecha_recepcion.isoformat() if rec.fecha_recepcion else "",
                f"{Decimal(rec.conformidad_pct or 0):.2f}",
                rec.get_estatus_display(),
                rec.observaciones or "",
            ]
        )
    return response


def _export_recepciones_xlsx(recepciones_qs) -> HttpResponse:
    recepciones = list(recepciones_qs)
    plan_ids = set()
    for rec in recepciones:
        solicitud_area = getattr(getattr(rec.orden, "solicitud", None), "area", "")
        plan_id = _extract_plan_id_from_scope(rec.orden.referencia) or _extract_plan_id_from_scope(solicitud_area)
        if plan_id:
            plan_ids.add(int(plan_id))
    planes_map = {
        p.id: p
        for p in PlanProduccion.objects.filter(id__in=plan_ids).only("id", "nombre", "notas")
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "recepciones_compra"
    ws.append(
        [
            "folio",
            "orden_folio",
            "origen",
            "plan_origen",
            "proveedor",
            "fecha_recepcion",
            "conformidad_pct",
            "estatus",
            "observaciones",
        ]
    )
    for rec in recepciones:
        source = _source_context_from_scope(
            area=getattr(getattr(rec.orden, "solicitud", None), "area", ""),
            referencia=rec.orden.referencia,
            planes_map=planes_map,
        )
        ws.append(
            [
                rec.folio,
                rec.orden.folio,
                source["source_tipo"],
                source["source_plan_nombre"] or source["source_label"],
                rec.orden.proveedor.nombre,
                rec.fecha_recepcion.isoformat() if rec.fecha_recepcion else "",
                float(Decimal(rec.conformidad_pct or 0)),
                rec.get_estatus_display(),
                rec.observaciones or "",
            ]
        )
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response["Content-Disposition"] = f'attachment; filename="recepciones_compra_{now_str}.xlsx"'
    return response


@login_required
def ordenes(request: HttpRequest) -> HttpResponse:
    if not can_view_compras(request.user):
        raise PermissionDenied("No tienes permisos para ver Compras.")

    if request.method == "POST":
        if not can_manage_compras(request.user):
            raise PermissionDenied("No tienes permisos para crear órdenes.")
        locked_plan_scope = _locked_plan_scope_from_request(request, current_view="ordenes")
        if (request.POST.get("source") or "").strip().lower() == "plan" and not locked_plan_scope:
            messages.error(request, "No se pudo resolver el plan activo para la orden.")
            return _redirect_scoped_list("compras:ordenes", request)
        proveedor_id = request.POST.get("proveedor_id")
        if proveedor_id:
            solicitud_raw = request.POST.get("solicitud_id")
            if not solicitud_raw:
                messages.error(request, "Debes seleccionar una solicitud aprobada para crear una orden.")
                return _redirect_scoped_list("compras:ordenes", request)

            solicitud = get_object_or_404(SolicitudCompra, pk=solicitud_raw)
            if solicitud.estatus != SolicitudCompra.STATUS_APROBADA:
                messages.error(request, f"La solicitud {solicitud.folio} no está aprobada.")
                return _redirect_scoped_list("compras:ordenes", request)
            if locked_plan_scope and not _scope_matches_plan_scope(
                plan_scope=locked_plan_scope["plan_scope"],
                area=solicitud.area,
            ):
                messages.error(
                    request,
                    f"La solicitud {solicitud.folio} no pertenece al plan {locked_plan_scope['label']}.",
                )
                return _redirect_scoped_list("compras:ordenes", request)

            orden = OrdenCompra.objects.create(
                proveedor_id=proveedor_id,
                solicitud=solicitud,
                referencia=locked_plan_scope["plan_scope"] if locked_plan_scope else f"SOLICITUD:{solicitud.folio}",
                fecha_emision=request.POST.get("fecha_emision") or None,
                fecha_entrega_estimada=request.POST.get("fecha_entrega_estimada") or None,
                monto_estimado=_to_decimal(request.POST.get("monto_estimado"), "0"),
                estatus=request.POST.get("estatus") or OrdenCompra.STATUS_BORRADOR,
            )
            log_event(
                request.user,
                "CREATE",
                "compras.OrdenCompra",
                orden.id,
                {"folio": orden.folio, "estatus": orden.estatus},
            )
        return _redirect_scoped_list("compras:ordenes", request)

    ordenes_qs = OrdenCompra.objects.select_related("proveedor", "solicitud").order_by("-creado_en")
    proveedor_filter = (request.GET.get("proveedor_id") or "").strip()
    source_filter = (request.GET.get("source") or "all").lower()
    if source_filter not in {"all", "manual", "plan", "reabasto_cedis"}:
        source_filter = "all"
    plan_filter = (request.GET.get("plan_id") or "").strip()
    estatus_filter = (request.GET.get("estatus") or "all").strip().upper() or "all"
    mes_filter = (request.GET.get("mes") or "").strip()
    q_filter = (request.GET.get("q") or "").strip()
    workflow_action_filter = (request.GET.get("workflow_action") or "all").strip().lower()
    blocker_key_filter = (request.GET.get("blocker_key") or "all").strip().lower()
    closure_key_raw = (request.GET.get("closure_key") or "all").strip().lower()
    handoff_key_raw = (request.GET.get("handoff_key") or "all").strip().lower()
    closure_key_filter = closure_key_raw
    handoff_key_filter = handoff_key_raw
    master_class_filter = (request.GET.get("master_class") or "all").strip()
    master_missing_filter = (request.GET.get("master_missing") or "all").strip()

    if proveedor_filter.isdigit():
        ordenes_qs = ordenes_qs.filter(proveedor_id=int(proveedor_filter))
    if source_filter == "reabasto_cedis":
        ordenes_qs = _filter_ordenes_by_scope(ordenes_qs, "plan", plan_filter)
    else:
        ordenes_qs = _filter_ordenes_by_scope(ordenes_qs, source_filter, plan_filter)
    if estatus_filter and estatus_filter not in {"ALL", "BLOCKED_ERP"}:
        ordenes_qs = ordenes_qs.filter(estatus=estatus_filter)
    parsed_month = _parse_month_filter(mes_filter)
    if parsed_month:
        y, m = parsed_month
        ordenes_qs = ordenes_qs.filter(fecha_emision__year=y, fecha_emision__month=m)
    if q_filter:
        ordenes_qs = ordenes_qs.filter(
            Q(folio__icontains=q_filter)
            | Q(referencia__icontains=q_filter)
            | Q(proveedor__nombre__icontains=q_filter)
            | Q(solicitud__folio__icontains=q_filter)
        )

    export_format = (request.GET.get("export") or "").strip().lower()
    if export_format == "csv":
        return _export_ordenes_csv(ordenes_qs)
    if export_format == "xlsx":
        return _export_ordenes_xlsx(ordenes_qs)

    query_without_export = request.GET.copy()
    query_without_export.pop("export", None)
    query_without_estatus = query_without_export.copy()
    query_without_estatus.pop("estatus", None)
    query_without_workflow = query_without_export.copy()
    query_without_workflow.pop("workflow_action", None)
    query_without_blocker = query_without_export.copy()
    query_without_blocker.pop("blocker_key", None)
    query_without_closure = query_without_export.copy()
    query_without_closure.pop("closure_key", None)
    query_without_handoff = query_without_export.copy()
    query_without_handoff.pop("handoff_key", None)
    plan_scope_context = _build_plan_scope_context(
        source_filter=source_filter,
        plan_filter=plan_filter,
        q_filter=q_filter,
        current_view="ordenes",
        closure_key_filter=closure_key_raw,
        handoff_key_filter=handoff_key_raw,
        master_class_filter=master_class_filter,
        master_missing_filter=master_missing_filter,
        session=request.session,
    )

    ordenes = list(ordenes_qs[:200])
    plan_ids = {
        int(plan_id)
        for plan_id in {
            _extract_plan_id_from_scope(orden.referencia)
            or _extract_plan_id_from_scope(getattr(orden.solicitud, "area", ""))
            for orden in ordenes
        }
        if plan_id
    }
    planes_map = {
        p.id: p
        for p in PlanProduccion.objects.filter(id__in=plan_ids).only("id", "nombre", "fecha_produccion", "notas")
    }
    recepciones_cerradas_por_orden = {
        row["orden_id"]: row["total"]
        for row in (
            RecepcionCompra.objects.filter(orden_id__in=[o.id for o in ordenes], estatus=RecepcionCompra.STATUS_CERRADA)
            .values("orden_id")
            .annotate(total=Count("id"))
        )
    } if ordenes else {}
    for orden in ordenes:
        orden.__dict__.update(
            _source_context_from_scope(
                area=getattr(orden.solicitud, "area", ""),
                referencia=orden.referencia,
                planes_map=planes_map,
            )
        )
        _enrich_orden_workflow(orden, int(recepciones_cerradas_por_orden.get(orden.id) or 0))
        orden.enterprise_master_blocker_details = (
            _enterprise_blocker_details_for_solicitud(orden.solicitud) if orden.solicitud_id else []
        )
    release_gate_rows = _ordenes_release_gate_rows(ordenes)
    if source_filter == "reabasto_cedis":
        ordenes = [orden for orden in ordenes if getattr(orden, "source_tipo", "") == "reabasto_cedis"]
    if estatus_filter == "BLOCKED_ERP":
        ordenes = [orden for orden in ordenes if orden.has_workflow_blockers]
    valid_workflow_actions = {
        "all",
        "corregir_datos",
        "enviar_proveedor",
        "confirmar_parcial",
        "registrar_recepcion",
        "completar_recepcion",
    }
    if workflow_action_filter in valid_workflow_actions and workflow_action_filter != "all":
        ordenes = [
            orden
            for orden in ordenes
            if getattr(orden, "workflow_action_code", "") == workflow_action_filter
        ]
    else:
        workflow_action_filter = "all"
    valid_blocker_keys = {"all", "sin_proveedor", "sin_emision", "monto_cero", "sin_entrega"}
    if blocker_key_filter in valid_blocker_keys and blocker_key_filter != "all":
        ordenes = [orden for orden in ordenes if _orden_has_blocker_key(orden, blocker_key_filter)]
    else:
        blocker_key_filter = "all"
    valid_closure_keys = {"all", "ordenes_sin_bloqueo"}
    if closure_key_filter in valid_closure_keys and closure_key_filter != "all":
        ordenes = [orden for orden in ordenes if getattr(orden, "has_workflow_blockers", False)]
    else:
        closure_key_filter = "all"
    valid_handoff_keys = {"all", "orden_recepcion"}
    if handoff_key_filter in valid_handoff_keys and handoff_key_filter != "all":
        ordenes = [
            orden
            for orden in ordenes
            if getattr(orden, "has_workflow_blockers", False)
            or orden.estatus == OrdenCompra.STATUS_CONFIRMADA
        ]
    else:
        handoff_key_filter = "all"
    ordenes = _filter_documents_by_master_blockers(ordenes, master_class_filter, master_missing_filter)

    solicitudes_form_qs = SolicitudCompra.objects.filter(estatus=SolicitudCompra.STATUS_APROBADA)
    if plan_scope_context:
        solicitudes_form_qs = solicitudes_form_qs.filter(area=plan_scope_context["plan_scope"])

    plan_ids_all = set()
    for scope_val in (
        list(OrdenCompra.objects.filter(referencia__startswith="PLAN_PRODUCCION:").values_list("referencia", flat=True).distinct()[:300])
        + list(OrdenCompra.objects.filter(solicitud__area__startswith="PLAN_PRODUCCION:").values_list("solicitud__area", flat=True).distinct()[:300])
    ):
        plan_id = _extract_plan_id_from_scope(scope_val)
        if plan_id:
            plan_ids_all.add(int(plan_id))
    plan_options = list(PlanProduccion.objects.filter(id__in=plan_ids_all).order_by("-fecha_produccion", "-id")[:100])

    context = {
        "ordenes": ordenes,
        "proveedores": Proveedor.objects.filter(activo=True).order_by("nombre")[:200],
        "solicitudes": solicitudes_form_qs.order_by("-creado_en")[:200],
        "status_choices": OrdenCompra.STATUS_CHOICES,
        "proveedor_filter": proveedor_filter,
        "source_filter": source_filter,
        "plan_filter": plan_filter,
        "plan_options": plan_options,
        "estatus_filter": estatus_filter,
        "mes_filter": mes_filter,
        "q_filter": q_filter,
        "workflow_action_filter": workflow_action_filter,
        "blocker_key_filter": blocker_key_filter,
        "closure_key_filter": closure_key_filter,
        "handoff_key_filter": handoff_key_filter,
        "master_class_filter": master_class_filter,
        "master_missing_filter": master_missing_filter,
        "master_class_choices": [
            ("all", "Todas las clases"),
            (Insumo.TIPO_MATERIA_PRIMA, "Materia prima"),
            (Insumo.TIPO_INTERNO, "Insumo interno"),
            (Insumo.TIPO_EMPAQUE, "Empaque"),
            ("unknown", "Sin catálogo"),
        ],
        "master_missing_choices": [
            ("all", "Todos los faltantes"),
            ("unidad", "Unidad base"),
            ("proveedor", "Proveedor principal"),
            ("categoria", "Categoría"),
            ("codigo_point", "Código Point"),
            ("other", "Otros"),
        ],
        "current_query": query_without_export.urlencode(),
        "current_query_without_estatus": query_without_estatus.urlencode(),
        "current_query_without_workflow": query_without_workflow.urlencode(),
        "current_query_without_blocker": query_without_blocker.urlencode(),
        "current_query_without_closure": query_without_closure.urlencode(),
        "current_query_without_handoff": query_without_handoff.urlencode(),
        "plan_scope_context": plan_scope_context,
        "can_manage_compras": can_manage_compras(request.user),
        "workflow_summary": _ordenes_workflow_summary(ordenes),
        "enterprise_board": _ordenes_enterprise_board(ordenes),
        "release_gate_rows": release_gate_rows,
        "executive_radar_rows": _document_executive_radar_rows(
            release_gate_rows,
            owner="Compras / Abastecimiento",
            fallback_url=reverse("compras:ordenes"),
        ),
        "trunk_handoff_rows": _document_trunk_handoff_rows(
            release_gate_rows,
            owner="Compras / Abastecimiento",
            fallback_url=reverse("compras:ordenes"),
        ),
        "erp_governance_rows": _document_governance_rows(
            release_gate_rows,
            owner="Compras / Abastecimiento",
        ),
        "critical_path_rows": _document_critical_path_rows(
            release_gate_rows,
            owner="Compras / Abastecimiento",
            fallback_url=reverse("compras:ordenes"),
        ),
        "erp_command_center": _document_command_center(
            release_gate_rows,
            owner="Compras / Abastecimiento",
            plan_scope_context=plan_scope_context,
            fallback_url=reverse("compras:ordenes"),
            fallback_cta="Abrir órdenes",
        ),
        "release_gate_completion": _document_release_gate_completion(release_gate_rows),
    }
    return render(request, "compras/ordenes.html", context)


@login_required
def recepciones(request: HttpRequest) -> HttpResponse:
    if not can_view_compras(request.user):
        raise PermissionDenied("No tienes permisos para ver Compras.")

    if request.method == "POST":
        if not can_manage_compras(request.user):
            raise PermissionDenied("No tienes permisos para registrar recepciones.")
        locked_plan_scope = _locked_plan_scope_from_request(request, current_view="recepciones")
        if (request.POST.get("source") or "").strip().lower() == "plan" and not locked_plan_scope:
            messages.error(request, "No se pudo resolver el plan activo para la recepción.")
            return _redirect_scoped_list("compras:recepciones", request)
        orden_id = request.POST.get("orden_id")
        if orden_id:
            orden = get_object_or_404(OrdenCompra, pk=orden_id)
            if orden.estatus in {OrdenCompra.STATUS_BORRADOR, OrdenCompra.STATUS_CERRADA}:
                messages.error(request, f"La orden {orden.folio} no admite recepciones en estatus {orden.get_estatus_display()}.")
                return _redirect_scoped_list("compras:recepciones", request)
            if locked_plan_scope and not _scope_matches_plan_scope(
                plan_scope=locked_plan_scope["plan_scope"],
                area=getattr(getattr(orden, "solicitud", None), "area", ""),
                referencia=orden.referencia,
            ):
                messages.error(
                    request,
                    f"La orden {orden.folio} no pertenece al plan {locked_plan_scope['label']}.",
                )
                return _redirect_scoped_list("compras:recepciones", request)

            recepcion = RecepcionCompra(
                orden=orden,
                fecha_recepcion=request.POST.get("fecha_recepcion") or None,
                conformidad_pct=_to_decimal(request.POST.get("conformidad_pct"), "100"),
                estatus=request.POST.get("estatus") or RecepcionCompra.STATUS_PENDIENTE,
                observaciones=request.POST.get("observaciones", "").strip(),
            )
            blockers = _enterprise_blockers_for_recepcion(recepcion)
            if blockers:
                messages.error(
                    request,
                    "No puedes registrar la recepción: " + "; ".join(blockers) + ".",
                )
                return _redirect_scoped_list("compras:recepciones", request)
            recepcion.save()
            log_event(
                request.user,
                "CREATE",
                "compras.RecepcionCompra",
                recepcion.id,
                {"folio": recepcion.folio, "estatus": recepcion.estatus},
            )
            if recepcion.estatus == RecepcionCompra.STATUS_CERRADA:
                _apply_recepcion_to_inventario(recepcion, acted_by=request.user)
                if orden.estatus != OrdenCompra.STATUS_CERRADA:
                    orden_prev = orden.estatus
                    orden.estatus = OrdenCompra.STATUS_CERRADA
                    orden.save(update_fields=["estatus"])
                    log_event(
                        request.user,
                    "APPROVE",
                    "compras.OrdenCompra",
                    orden.id,
                    {"from": orden_prev, "to": OrdenCompra.STATUS_CERRADA, "folio": orden.folio, "source": recepcion.folio},
                )
        return _redirect_scoped_list("compras:recepciones", request)

    recepciones_qs = RecepcionCompra.objects.select_related("orden", "orden__proveedor").order_by("-creado_en")
    proveedor_filter = (request.GET.get("proveedor_id") or "").strip()
    source_filter = (request.GET.get("source") or "all").lower()
    if source_filter not in {"all", "manual", "plan", "reabasto_cedis"}:
        source_filter = "all"
    plan_filter = (request.GET.get("plan_id") or "").strip()
    estatus_filter = (request.GET.get("estatus") or "all").strip().upper() or "all"
    mes_filter = (request.GET.get("mes") or "").strip()
    q_filter = (request.GET.get("q") or "").strip()
    workflow_action_filter = (request.GET.get("workflow_action") or "all").strip().lower()
    blocker_key_filter = (request.GET.get("blocker_key") or "all").strip().lower()
    closure_key_raw = (request.GET.get("closure_key") or "all").strip().lower()
    handoff_key_raw = (request.GET.get("handoff_key") or "all").strip().lower()
    closure_key_filter = closure_key_raw
    handoff_key_filter = handoff_key_raw
    master_class_filter = (request.GET.get("master_class") or "all").strip()
    master_missing_filter = (request.GET.get("master_missing") or "all").strip()

    if proveedor_filter.isdigit():
        recepciones_qs = recepciones_qs.filter(orden__proveedor_id=int(proveedor_filter))
    if source_filter == "reabasto_cedis":
        recepciones_qs = _filter_recepciones_by_scope(recepciones_qs, "plan", plan_filter)
    else:
        recepciones_qs = _filter_recepciones_by_scope(recepciones_qs, source_filter, plan_filter)
    if estatus_filter and estatus_filter not in {"ALL", "BLOCKED_ERP"}:
        recepciones_qs = recepciones_qs.filter(estatus=estatus_filter)
    parsed_month = _parse_month_filter(mes_filter)
    if parsed_month:
        y, m = parsed_month
        recepciones_qs = recepciones_qs.filter(fecha_recepcion__year=y, fecha_recepcion__month=m)
    if q_filter:
        recepciones_qs = recepciones_qs.filter(
            Q(folio__icontains=q_filter)
            | Q(orden__folio__icontains=q_filter)
            | Q(orden__referencia__icontains=q_filter)
            | Q(orden__proveedor__nombre__icontains=q_filter)
            | Q(observaciones__icontains=q_filter)
        )

    export_format = (request.GET.get("export") or "").strip().lower()
    if export_format == "csv":
        return _export_recepciones_csv(recepciones_qs)
    if export_format == "xlsx":
        return _export_recepciones_xlsx(recepciones_qs)

    query_without_export = request.GET.copy()
    query_without_export.pop("export", None)
    query_without_estatus = query_without_export.copy()
    query_without_estatus.pop("estatus", None)
    query_without_workflow = query_without_export.copy()
    query_without_workflow.pop("workflow_action", None)
    query_without_blocker = query_without_export.copy()
    query_without_blocker.pop("blocker_key", None)
    query_without_closure = query_without_export.copy()
    query_without_closure.pop("closure_key", None)
    query_without_handoff = query_without_export.copy()
    query_without_handoff.pop("handoff_key", None)
    plan_scope_context = _build_plan_scope_context(
        source_filter=source_filter,
        plan_filter=plan_filter,
        q_filter=q_filter,
        current_view="recepciones",
        closure_key_filter=closure_key_raw,
        handoff_key_filter=handoff_key_raw,
        master_class_filter=master_class_filter,
        master_missing_filter=master_missing_filter,
        session=request.session,
    )

    recepciones = list(recepciones_qs[:200])
    plan_ids: set[int] = set()
    for recepcion in recepciones:
        solicitud_area = getattr(getattr(recepcion.orden, "solicitud", None), "area", "")
        plan_id = _extract_plan_id_from_scope(recepcion.orden.referencia) or _extract_plan_id_from_scope(solicitud_area)
        if plan_id:
            plan_ids.add(int(plan_id))
    planes_map = {
        p.id: p
        for p in PlanProduccion.objects.filter(id__in=plan_ids).only("id", "nombre", "fecha_produccion", "notas")
    }
    for recepcion in recepciones:
        recepcion.__dict__.update(
            _source_context_from_scope(
                area=getattr(getattr(recepcion.orden, "solicitud", None), "area", ""),
                referencia=recepcion.orden.referencia,
                planes_map=planes_map,
            )
        )
        _enrich_recepcion_workflow(recepcion)
        recepcion.enterprise_master_blocker_details = (
            _enterprise_blocker_details_for_solicitud(recepcion.orden.solicitud)
            if recepcion.orden_id and getattr(recepcion.orden, "solicitud_id", None)
            else []
        )
    release_gate_rows = _recepciones_release_gate_rows(recepciones)
    if source_filter == "reabasto_cedis":
        recepciones = [recepcion for recepcion in recepciones if getattr(recepcion, "source_tipo", "") == "reabasto_cedis"]
    if estatus_filter == "BLOCKED_ERP":
        recepciones = [recepcion for recepcion in recepciones if recepcion.has_workflow_blockers]
    valid_workflow_actions = {"all", "corregir_recepcion", "cerrar_diferencias", "cerrar_aplicar"}
    if workflow_action_filter in valid_workflow_actions and workflow_action_filter != "all":
        recepciones = [
            recepcion
            for recepcion in recepciones
            if getattr(recepcion, "workflow_action_code", "") == workflow_action_filter
        ]
    else:
        workflow_action_filter = "all"
    valid_blocker_keys = {"all", "sin_fecha", "conformidad", "sin_observacion", "sin_justificacion"}
    if blocker_key_filter in valid_blocker_keys and blocker_key_filter != "all":
        recepciones = [
            recepcion
            for recepcion in recepciones
            if _recepcion_has_blocker_key(recepcion, blocker_key_filter)
        ]
    else:
        blocker_key_filter = "all"
    valid_closure_keys = {"all", "recepciones_aplicadas"}
    if closure_key_filter in valid_closure_keys and closure_key_filter != "all":
        recepciones = [
            recepcion
            for recepcion in recepciones
            if getattr(recepcion, "has_workflow_blockers", False)
            or recepcion.estatus in {RecepcionCompra.STATUS_DIFERENCIAS, RecepcionCompra.STATUS_PENDIENTE}
        ]
    else:
        closure_key_filter = "all"
    valid_handoff_keys = {"all", "recepcion_cierre"}
    if handoff_key_filter in valid_handoff_keys and handoff_key_filter != "all":
        recepciones = [
            recepcion
            for recepcion in recepciones
            if getattr(recepcion, "has_workflow_blockers", False)
            or recepcion.estatus in {RecepcionCompra.STATUS_DIFERENCIAS, RecepcionCompra.STATUS_PENDIENTE}
        ]
    else:
        handoff_key_filter = "all"
    recepciones = _filter_documents_by_master_blockers(recepciones, master_class_filter, master_missing_filter)

    ordenes_form_qs = (
        OrdenCompra.objects.select_related("proveedor")
        .exclude(estatus=OrdenCompra.STATUS_BORRADOR)
        .exclude(estatus=OrdenCompra.STATUS_CERRADA)
    )
    if plan_scope_context:
        ordenes_form_qs = ordenes_form_qs.filter(referencia=plan_scope_context["plan_scope"])

    plan_ids_all = set()
    for scope_val in (
        list(RecepcionCompra.objects.filter(orden__referencia__startswith="PLAN_PRODUCCION:").values_list("orden__referencia", flat=True).distinct()[:300])
        + list(RecepcionCompra.objects.filter(orden__solicitud__area__startswith="PLAN_PRODUCCION:").values_list("orden__solicitud__area", flat=True).distinct()[:300])
    ):
        plan_id = _extract_plan_id_from_scope(scope_val)
        if plan_id:
            plan_ids_all.add(int(plan_id))
    plan_options = list(PlanProduccion.objects.filter(id__in=plan_ids_all).order_by("-fecha_produccion", "-id")[:100])

    context = {
        "recepciones": recepciones,
        "ordenes": ordenes_form_qs.order_by("-creado_en")[:200],
        "status_choices": RecepcionCompra.STATUS_CHOICES,
        "proveedores": Proveedor.objects.filter(activo=True).order_by("nombre")[:200],
        "proveedor_filter": proveedor_filter,
        "source_filter": source_filter,
        "plan_filter": plan_filter,
        "plan_options": plan_options,
        "estatus_filter": estatus_filter,
        "mes_filter": mes_filter,
        "q_filter": q_filter,
        "workflow_action_filter": workflow_action_filter,
        "blocker_key_filter": blocker_key_filter,
        "closure_key_filter": closure_key_filter,
        "handoff_key_filter": handoff_key_filter,
        "master_class_filter": master_class_filter,
        "master_missing_filter": master_missing_filter,
        "master_class_choices": [
            ("all", "Todas las clases"),
            (Insumo.TIPO_MATERIA_PRIMA, "Materia prima"),
            (Insumo.TIPO_INTERNO, "Insumo interno"),
            (Insumo.TIPO_EMPAQUE, "Empaque"),
            ("unknown", "Sin catálogo"),
        ],
        "master_missing_choices": [
            ("all", "Todos los faltantes"),
            ("unidad", "Unidad base"),
            ("proveedor", "Proveedor principal"),
            ("categoria", "Categoría"),
            ("codigo_point", "Código Point"),
            ("other", "Otros"),
        ],
        "current_query": query_without_export.urlencode(),
        "current_query_without_estatus": query_without_estatus.urlencode(),
        "current_query_without_workflow": query_without_workflow.urlencode(),
        "current_query_without_blocker": query_without_blocker.urlencode(),
        "current_query_without_closure": query_without_closure.urlencode(),
        "current_query_without_handoff": query_without_handoff.urlencode(),
        "plan_scope_context": plan_scope_context,
        "can_manage_compras": can_manage_compras(request.user),
        "workflow_summary": _recepciones_workflow_summary(recepciones),
        "enterprise_board": _recepciones_enterprise_board(recepciones),
        "release_gate_rows": release_gate_rows,
        "executive_radar_rows": _document_executive_radar_rows(
            release_gate_rows,
            owner="Compras / Recepción",
            fallback_url=reverse("compras:recepciones"),
        ),
        "trunk_handoff_rows": _document_trunk_handoff_rows(
            release_gate_rows,
            owner="Compras / Recepción",
            fallback_url=reverse("compras:recepciones"),
        ),
        "erp_governance_rows": _document_governance_rows(
            release_gate_rows,
            owner="Compras / Recepción",
        ),
        "critical_path_rows": _document_critical_path_rows(
            release_gate_rows,
            owner="Compras / Recepción",
            fallback_url=reverse("compras:recepciones"),
        ),
        "erp_command_center": _document_command_center(
            release_gate_rows,
            owner="Compras / Recepción",
            plan_scope_context=plan_scope_context,
            fallback_url=reverse("compras:recepciones"),
            fallback_cta="Abrir recepciones",
        ),
        "release_gate_completion": _document_release_gate_completion(release_gate_rows),
    }
    return render(request, "compras/recepciones.html", context)


@login_required
@require_POST
def actualizar_solicitud_estatus(request: HttpRequest, pk: int, estatus: str) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para aprobar/rechazar solicitudes.")

    solicitud = get_object_or_404(SolicitudCompra, pk=pk)
    prev = solicitud.estatus
    blockers = _enterprise_blockers_for_solicitud(solicitud)
    if blockers and estatus in {SolicitudCompra.STATUS_EN_REVISION, SolicitudCompra.STATUS_APROBADA}:
        messages.error(
            request,
            f"No puedes mover {solicitud.folio} a {estatus}: " + "; ".join(blockers) + ".",
        )
        return _redirect_scoped_list("compras:solicitudes", request, preserve_query=True)
    if _can_transition_solicitud(prev, estatus):
        solicitud.estatus = estatus
        solicitud.save(update_fields=["estatus"])
        log_event(
            request.user,
            "APPROVE",
            "compras.SolicitudCompra",
            solicitud.id,
            {"from": prev, "to": estatus, "folio": solicitud.folio},
        )
    return _redirect_scoped_list("compras:solicitudes", request, preserve_query=True)


@login_required
@require_POST
def eliminar_solicitud(request: HttpRequest, pk: int) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para eliminar solicitudes.")

    solicitud = get_object_or_404(SolicitudCompra, pk=pk)
    has_open_order = OrdenCompra.objects.filter(solicitud=solicitud).exclude(estatus=OrdenCompra.STATUS_CERRADA).exists()
    return_query = (request.POST.get("return_query") or "").strip()
    if has_open_order:
        messages.error(request, f"No puedes eliminar {solicitud.folio}: tiene una orden de compra activa.")
    else:
        solicitud_id = solicitud.id
        folio = solicitud.folio
        solicitud.delete()
        log_event(
            request.user,
            "DELETE",
            "compras.SolicitudCompra",
            solicitud_id,
            {"folio": folio},
        )
        messages.success(request, f"Solicitud {folio} eliminada.")

    if return_query:
        return redirect(f"{reverse('compras:solicitudes')}?{return_query}")
    return redirect("compras:solicitudes")


@login_required
@require_POST
def actualizar_orden_estatus(request: HttpRequest, pk: int, estatus: str) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para operar órdenes.")

    orden = get_object_or_404(OrdenCompra, pk=pk)
    prev = orden.estatus
    blockers = _enterprise_blockers_for_orden(orden)

    if estatus in {OrdenCompra.STATUS_ENVIADA, OrdenCompra.STATUS_CONFIRMADA, OrdenCompra.STATUS_PARCIAL} and blockers:
        messages.error(
            request,
            f"No puedes mover {orden.folio} a {estatus}: " + "; ".join(blockers) + ".",
        )
        return _redirect_scoped_list("compras:ordenes", request, preserve_query=True)

    if estatus == OrdenCompra.STATUS_CERRADA:
        has_closed_recepcion = RecepcionCompra.objects.filter(
            orden=orden,
            estatus=RecepcionCompra.STATUS_CERRADA,
        ).exists()
        if not has_closed_recepcion:
            messages.error(request, f"No puedes cerrar {orden.folio} sin al menos una recepción cerrada.")
            return _redirect_scoped_list("compras:ordenes", request, preserve_query=True)

    if _can_transition_orden(prev, estatus):
        orden.estatus = estatus
        orden.save(update_fields=["estatus"])
        log_event(
            request.user,
            "APPROVE",
            "compras.OrdenCompra",
            orden.id,
            {"from": prev, "to": estatus, "folio": orden.folio},
        )
    return _redirect_scoped_list("compras:ordenes", request, preserve_query=True)


@login_required
@require_POST
def actualizar_recepcion_estatus(request: HttpRequest, pk: int, estatus: str) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para cerrar recepciones.")

    recepcion = get_object_or_404(RecepcionCompra, pk=pk)
    prev = recepcion.estatus
    recepcion_preview = RecepcionCompra(
        orden=recepcion.orden,
        fecha_recepcion=recepcion.fecha_recepcion,
        conformidad_pct=recepcion.conformidad_pct,
        estatus=estatus,
        observaciones=recepcion.observaciones,
    )
    blockers = _enterprise_blockers_for_recepcion(recepcion_preview)
    if estatus in {RecepcionCompra.STATUS_DIFERENCIAS, RecepcionCompra.STATUS_CERRADA} and blockers:
        messages.error(
            request,
            f"No puedes mover {recepcion.folio} a {estatus}: " + "; ".join(blockers) + ".",
        )
        return _redirect_scoped_list("compras:recepciones", request, preserve_query=True)
    if _can_transition_recepcion(prev, estatus):
        recepcion.estatus = estatus
        recepcion.save(update_fields=["estatus"])
        log_event(
            request.user,
            "APPROVE",
            "compras.RecepcionCompra",
            recepcion.id,
            {"from": prev, "to": estatus, "folio": recepcion.folio},
        )

        # Si la recepción quedó cerrada, marcamos la orden cerrada automáticamente.
        if estatus == RecepcionCompra.STATUS_CERRADA:
            _apply_recepcion_to_inventario(recepcion, acted_by=request.user)
            if recepcion.orden.estatus != OrdenCompra.STATUS_CERRADA:
                orden_prev = recepcion.orden.estatus
                recepcion.orden.estatus = OrdenCompra.STATUS_CERRADA
                recepcion.orden.save(update_fields=["estatus"])
                log_event(
                    request.user,
                    "APPROVE",
                    "compras.OrdenCompra",
                    recepcion.orden.id,
                    {"from": orden_prev, "to": OrdenCompra.STATUS_CERRADA, "folio": recepcion.orden.folio, "source": recepcion.folio},
                )
    return _redirect_scoped_list("compras:recepciones", request, preserve_query=True)


@login_required
@require_POST
def crear_orden_desde_solicitud(request: HttpRequest, pk: int) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para crear órdenes.")

    solicitud = get_object_or_404(SolicitudCompra.objects.select_related("insumo", "proveedor_sugerido"), pk=pk)
    locked_plan_scope = _locked_plan_scope_from_request(request, current_view="solicitudes")
    if (request.POST.get("source") or "").strip().lower() == "plan" and not locked_plan_scope:
        messages.error(request, "No se pudo resolver el plan activo para crear la orden.")
        return _redirect_scoped_list("compras:solicitudes", request, preserve_query=True)
    if locked_plan_scope and not _scope_matches_plan_scope(
        plan_scope=locked_plan_scope["plan_scope"],
        area=solicitud.area,
    ):
        messages.error(
            request,
            f"La solicitud {solicitud.folio} no pertenece al plan {locked_plan_scope['label']}.",
        )
        return _redirect_scoped_list("compras:solicitudes", request, preserve_query=True)
    if solicitud.estatus != SolicitudCompra.STATUS_APROBADA:
        messages.error(request, f"La solicitud {solicitud.folio} no está aprobada.")
        return _redirect_scoped_list("compras:solicitudes", request, preserve_query=True)

    has_open_order = OrdenCompra.objects.filter(solicitud=solicitud).exclude(estatus=OrdenCompra.STATUS_CERRADA).exists()
    if has_open_order:
        messages.info(request, f"La solicitud {solicitud.folio} ya tiene una orden activa.")
        return _redirect_scoped_list("compras:ordenes", request, preserve_query=True)

    if locked_plan_scope:
        master_demand_gate = locked_plan_scope.get("master_demand_gate") or {}
        if (
            str(master_demand_gate.get("tone") or "") == "danger"
            and int(master_demand_gate.get("blockers") or 0) > 0
        ):
            messages.error(
                request,
                f"No puedes emitir la orden de compra desde {solicitud.folio}: el plan tiene demanda crítica bloqueada por maestro.",
            )
            return _redirect_scoped_list("compras:solicitudes", request, preserve_query=True)

    insumo_canonical = canonical_insumo_by_id(solicitud.insumo_id) or solicitud.insumo
    if insumo_canonical and solicitud.insumo_id != insumo_canonical.id:
        previous_insumo_id = solicitud.insumo_id
        previous_insumo_nombre = getattr(solicitud.insumo, "nombre", "") if getattr(solicitud, "insumo", None) else ""
        solicitud.insumo = insumo_canonical
        solicitud.save(update_fields=["insumo"])
        log_event(
            request.user,
            "UPDATE",
            "compras.SolicitudCompra",
            solicitud.id,
            {
                "folio": solicitud.folio,
                "field": "insumo",
                "from_id": previous_insumo_id,
                "from_nombre": previous_insumo_nombre,
                "to_id": insumo_canonical.id,
                "to_nombre": insumo_canonical.nombre,
                "source": "crear_orden_desde_solicitud",
            },
        )
        solicitud = SolicitudCompra.objects.select_related("insumo", "proveedor_sugerido").get(pk=solicitud.pk)

    blockers = _enterprise_blockers_for_solicitud(solicitud)
    if blockers:
        messages.error(
            request,
            f"No puedes emitir la orden de compra desde {solicitud.folio}: " + "; ".join(blockers) + ".",
        )
        return _redirect_scoped_list("compras:solicitudes", request, preserve_query=True)

    insumo_canonical = canonical_insumo_by_id(solicitud.insumo_id) or solicitud.insumo

    proveedor = solicitud.proveedor_sugerido or insumo_canonical.proveedor_principal or solicitud.insumo.proveedor_principal
    if not proveedor:
        messages.error(request, f"La solicitud {solicitud.folio} no tiene proveedor sugerido. Asigna uno y reintenta.")
        return _redirect_scoped_list("compras:solicitudes", request, preserve_query=True)

    latest_cost = latest_costo_canonico(insumo_id=insumo_canonical.id)
    monto_estimado = (solicitud.cantidad or Decimal("0")) * (latest_cost if latest_cost else Decimal("0"))

    orden = OrdenCompra.objects.create(
        solicitud=solicitud,
        proveedor=proveedor,
        referencia=locked_plan_scope["plan_scope"] if locked_plan_scope else f"SOLICITUD:{solicitud.folio}",
        fecha_emision=timezone.localdate(),
        fecha_entrega_estimada=solicitud.fecha_requerida,
        monto_estimado=monto_estimado,
        estatus=OrdenCompra.STATUS_BORRADOR,
    )
    log_event(
        request.user,
        "CREATE",
        "compras.OrdenCompra",
        orden.id,
        {"folio": orden.folio, "estatus": orden.estatus, "source": f"solicitud:{solicitud.folio}"},
    )
    messages.success(request, f"Orden {orden.folio} creada desde solicitud {solicitud.folio}.")
    return _redirect_scoped_list("compras:ordenes", request, preserve_query=True)
