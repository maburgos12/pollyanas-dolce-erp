from io import BytesIO
from datetime import date, datetime, timedelta
from decimal import Decimal
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from core.models import Sucursal
from compras.models import (
    OrdenCompra,
    PresupuestoCompraCategoria,
    PresupuestoCompraPeriodo,
    PresupuestoCompraProveedor,
    RecepcionCompra,
    SolicitudCompra,
)
from inventario.models import ExistenciaInsumo, MovimientoInventario
from maestros.models import CostoInsumo, Insumo, Proveedor, UnidadMedida
from recetas.models import (
    LineaReceta,
    PlanProduccion,
    PlanProduccionItem,
    Receta,
    SolicitudVenta,
    VentaHistorica,
)


class ComprasFase2FiltersTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_test",
            email="admin_test@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

        self.proveedor = Proveedor.objects.create(nombre="Proveedor Test", activo=True)
        self.unidad_kg = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.unidad_lt = UnidadMedida.objects.create(
            codigo="lt",
            nombre="Litro",
            tipo=UnidadMedida.TIPO_VOLUMEN,
            factor_to_base=Decimal("1000"),
        )

        self.insumo_masa_blank = Insumo.objects.create(
            nombre="Harina sin categoria",
            categoria="",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        self.insumo_masa_explicit = Insumo.objects.create(
            nombre="Mantequilla categoria masa",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        self.insumo_volumen = Insumo.objects.create(
            nombre="Leche sin categoria",
            categoria="",
            unidad_base=self.unidad_lt,
            proveedor_principal=self.proveedor,
            activo=True,
        )

        CostoInsumo.objects.create(
            insumo=self.insumo_masa_blank,
            proveedor=self.proveedor,
            costo_unitario=Decimal("10"),
            source_hash="cost-harina-1",
        )
        CostoInsumo.objects.create(
            insumo=self.insumo_masa_explicit,
            proveedor=self.proveedor,
            costo_unitario=Decimal("5"),
            source_hash="cost-mantequilla-1",
        )
        CostoInsumo.objects.create(
            insumo=self.insumo_volumen,
            proveedor=self.proveedor,
            costo_unitario=Decimal("7"),
            source_hash="cost-leche-1",
        )

        self.periodo_mes = "2026-02"
        self.fecha_base = date(2026, 2, 10)
        self.solicitud_masa_blank = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=self.insumo_masa_blank,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("2"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        self.solicitud_masa_explicit = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=self.insumo_masa_explicit,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("1"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        self.solicitud_volumen = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=self.insumo_volumen,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("3"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_APROBADA,
        )

        OrdenCompra.objects.create(
            solicitud=self.solicitud_masa_blank,
            proveedor=self.proveedor,
            fecha_emision=self.fecha_base,
            monto_estimado=Decimal("30"),
            estatus=OrdenCompra.STATUS_ENVIADA,
        )
        OrdenCompra.objects.create(
            solicitud=self.solicitud_volumen,
            proveedor=self.proveedor,
            fecha_emision=self.fecha_base,
            monto_estimado=Decimal("100"),
            estatus=OrdenCompra.STATUS_ENVIADA,
        )

        self.receta_plan = Receta.objects.create(
            nombre="Base prueba plan",
            hash_contenido="test-hash-plan-001",
        )
        LineaReceta.objects.create(
            receta=self.receta_plan,
            posicion=1,
            insumo=self.insumo_masa_blank,
            insumo_texto="Harina",
            cantidad=Decimal("2"),
            unidad=self.unidad_kg,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("10"),
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100,
            match_status=LineaReceta.STATUS_AUTO,
        )
        self.plan = PlanProduccion.objects.create(
            nombre="Plan Febrero Test",
            fecha_produccion=self.fecha_base,
        )
        PlanProduccionItem.objects.create(
            plan=self.plan,
            receta=self.receta_plan,
            cantidad=Decimal("1"),
        )
        MovimientoInventario.objects.create(
            fecha=timezone.make_aware(datetime(2026, 2, 10, 11, 0, 0)),
            tipo=MovimientoInventario.TIPO_CONSUMO,
            insumo=self.insumo_masa_blank,
            cantidad=Decimal("3"),
            referencia=f"PLAN_PRODUCCION:{self.plan.id}",
        )

    def test_resumen_api_aplica_filtro_categoria(self):
        url = reverse("compras:solicitudes_resumen_api")
        response = self.client.get(
            url,
            {
                "periodo_tipo": "mes",
                "periodo_mes": self.periodo_mes,
                "categoria": "Masa",
            },
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()

        self.assertEqual(payload["filters"]["categoria"], "Masa")
        self.assertEqual(payload["totals"]["solicitudes_count"], 2)
        self.assertAlmostEqual(payload["totals"]["presupuesto_estimado_total"], 25.0, places=2)
        self.assertAlmostEqual(payload["totals"]["presupuesto_ejecutado_total"], 30.0, places=2)

        categorias = {row["categoria"]: row for row in payload["top_categorias"]}
        self.assertIn("Masa", categorias)
        self.assertAlmostEqual(categorias["Masa"]["estimado"], 25.0, places=2)
        self.assertAlmostEqual(categorias["Masa"]["ejecutado"], 30.0, places=2)

    def test_solicitudes_view_context_preserva_categoria(self):
        url = reverse("compras:solicitudes")
        response = self.client.get(
            url,
            {
                "periodo_tipo": "mes",
                "periodo_mes": self.periodo_mes,
                "categoria": "Masa",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["categoria_filter"], "Masa")
        self.assertEqual(len(response.context["solicitudes"]), 2)
        self.assertIn("categoria=Masa", response.context["current_query"])

    def test_solicitudes_view_exposes_plan_scope_context(self):
        SolicitudCompra.objects.create(
            area=f"PLAN_PRODUCCION:{self.plan.id}",
            solicitante="planificador",
            insumo=self.insumo_masa_blank,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("1"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        response = self.client.get(
            reverse("compras:solicitudes"),
            {"source": "plan", "plan_id": str(self.plan.id)},
        )
        self.assertEqual(response.status_code, 200)
        ctx = response.context["plan_scope_context"]
        self.assertIsNotNone(ctx)
        self.assertEqual(ctx["plan_id"], self.plan.id)
        self.assertEqual(ctx["summary_label"], "Con bloqueos")
        self.assertEqual(ctx["blocked_total"], 1)
        self.assertEqual(ctx["stage_label"], "Validación de solicitudes")
        self.assertIn("BLOCKED_ERP", ctx["next_action"]["url"])
        self.assertEqual([item["label"] for item in ctx["pipeline_steps"]], ["Solicitudes", "Órdenes", "Recepciones"])
        self.assertEqual(ctx["pipeline_steps"][0]["semaphore_label"], "Rojo")
        self.assertEqual(ctx["pipeline_steps"][0]["blocked"], 1)
        self.assertTrue(ctx["pipeline_steps"][0]["is_active"])
        self.assertIn("master_demand_gate", ctx)
        self.assertEqual(ctx["pipeline_steps"][0]["action_label"], "Resolver bloqueos")
        self.assertIn("habilitar emisión de órdenes", ctx["pipeline_steps"][0]["action_detail"])
        self.assertTrue(ctx["master_focus_rows"])
        focus_row = ctx["master_focus_rows"][0]
        self.assertIn(f"insumo_id={self.insumo_masa_blank.id}", focus_row["action_url"])
        self.assertEqual(focus_row["edit_url"], reverse("maestros:insumo_update", args=[self.insumo_masa_blank.id]))
        self.assertContains(response, "Editar artículo")
        self.assertEqual(ctx["stage_focus"]["label"], "Solicitudes")
        self.assertEqual(len(ctx["stage_focus"]["rows"]), 1)
        self.assertEqual(len(ctx["closure_checks"]), 3)
        self.assertEqual(ctx["closure_checks"][0]["action_label"], "Revisar solicitudes bloqueadas")
        self.assertIn("closure_key=solicitudes_liberadas", ctx["closure_checks"][0]["action_url"])
        self.assertIn("Corrige bloqueos ERP", ctx["closure_checks"][0]["action_detail"])
        self.assertEqual(ctx["closure_focus"]["label"], "Solicitudes liberadas")
        self.assertEqual(ctx["closure_focus"]["tone"], "danger")
        self.assertGreaterEqual(len(ctx["closure_focus_rows"]), 1)
        self.assertEqual(ctx["closure_focus_rows"][0]["scope"], "Solicitud")
        self.assertEqual([item["label"] for item in ctx["handoff_checks"]], ["Solicitud → Orden", "Orden → Recepción", "Recepción → Cierre"])
        self.assertEqual(ctx["handoff_checks"][0]["action_label"], "Resolver solicitudes bloqueadas")
        self.assertIn("handoff_key=solicitud_orden", ctx["handoff_checks"][0]["action_url"])
        self.assertEqual(ctx["handoff_focus"]["label"], "Solicitud → Orden")
        self.assertEqual(ctx["handoff_focus"]["tone"], "danger")
        self.assertGreaterEqual(len(ctx["handoff_focus_rows"]), 1)
        self.assertEqual(ctx["handoff_focus_rows"][0]["scope"], "Solicitud")
        self.assertIn("master_focus", ctx)
        self.assertEqual(ctx["master_focus"]["class_label"], "Materia prima")
        self.assertGreaterEqual(len(ctx["master_focus_rows"]), 1)

    def test_solicitudes_view_exposes_plan_demand_signal(self):
        VentaHistorica.objects.create(
            receta=self.receta_plan,
            fecha=date(2026, 2, 1),
            cantidad=Decimal("12"),
            tickets=3,
            monto_total=Decimal("360"),
        )
        VentaHistorica.objects.create(
            receta=self.receta_plan,
            fecha=date(2026, 2, 3),
            cantidad=Decimal("8"),
            tickets=2,
            monto_total=Decimal("240"),
        )
        SolicitudVenta.objects.create(
            receta=self.receta_plan,
            alcance=SolicitudVenta.ALCANCE_MES,
            periodo="2026-02",
            fecha_inicio=date(2026, 2, 1),
            fecha_fin=date(2026, 2, 28),
            cantidad=Decimal("10"),
        )
        session = self.client.session
        session["pronostico_estadistico_preview"] = {
            "alcance": "mes",
            "periodo": "2026-02",
            "target_start": "2026-02-01",
            "target_end": "2026-02-28",
            "sucursal_id": None,
            "sucursal_nombre": "Todas",
            "rows": [
                {
                    "receta_id": self.receta_plan.id,
                    "receta": self.receta_plan.nombre,
                    "forecast_qty": 10.0,
                    "forecast_low": 9.0,
                    "forecast_high": 11.0,
                    "desviacion": 1.0,
                    "muestras": 2,
                    "pronostico_actual": 10.0,
                    "delta": 0.0,
                    "recomendacion": "MANTENER",
                    "observaciones": 2,
                    "confianza": 82.5,
                }
            ],
            "totals": {
                "recetas_count": 1,
                "forecast_total": 10.0,
                "forecast_low_total": 9.0,
                "forecast_high_total": 11.0,
                "pronostico_total": 10.0,
                "delta_total": 0.0,
            },
        }
        session.save()
        sucursal = Sucursal.objects.create(codigo="SUC-PLAN-CRIT", nombre="Sucursal Plan Critica")
        VentaHistorica.objects.create(
            receta=self.receta_plan,
            sucursal=sucursal,
            fecha=self.fecha_base - timedelta(days=2),
            cantidad=Decimal("95"),
            tickets=8,
        )

        response = self.client.get(
            reverse("compras:solicitudes"),
            {"source": "plan", "plan_id": str(self.plan.id)},
        )
        self.assertEqual(response.status_code, 200)
        ctx = response.context["plan_scope_context"]
        signal = response.context["plan_scope_context"]["demand_signal"]
        self.assertIsNotNone(signal)
        self.assertEqual(signal["historico_days"], 3)
        self.assertEqual(signal["historico_years"], 1)
        self.assertEqual(signal["comparable_years"], 1)
        self.assertEqual(signal["forecast_count"], 1)
        self.assertEqual(signal["alignment_pct"], 100)
        self.assertIn("commercial_priority_rows", ctx)
        self.assertTrue(ctx["commercial_priority_rows"])
        self.assertEqual(ctx["master_demand_gate"]["tone"], "danger")
        self.assertTrue(ctx["master_demand_blocks_issue"])
        self.assertEqual(ctx["stage_label"], "Demanda crítica bloqueada")
        self.assertEqual(ctx["stage_tone"], "danger")
        self.assertIn("maestros/insumos", ctx["next_action"]["url"])
        self.assertTrue(ctx["critical_master_demand_rows"])
        self.assertIn("daily_critical_close_focus", ctx)
        self.assertIsNotNone(ctx["daily_critical_close_focus"])
        self.assertContains(response, "Señal histórica de demanda del plan")
        self.assertContains(response, "Años observados")
        self.assertContains(response, "Temporadas comparables")
        self.assertContains(response, "Control de demanda comercial")
        self.assertContains(response, "Control de maestro crítico por demanda")
        self.assertContains(response, "Cierre prioritario del día")
        self.assertContains(response, "Liberación documental retenida")
        self.assertContains(response, "Demanda crítica bloqueada del plan")
        self.assertContains(response, "Sucursales que empujan el plan")
        self.assertContains(response, "Insumo a asegurar por sucursal")
        self.assertContains(response, "Artículos prioritarios por demanda del plan")
        self.assertContains(response, "Faltante maestro")
        self.assertContains(response, self.insumo_masa_blank.nombre)
        self.assertContains(response, self.receta_plan.nombre)
        self.assertContains(response, "Alineación forecast/solicitud")
        self.assertContains(response, self.plan.nombre)
        self.assertContains(response, f"plan_id={self.plan.id}")
        self.assertContains(response, "Bloqueos del plan")
        self.assertContains(response, "Etapa documental actual")
        self.assertContains(response, "Demanda crítica bloqueada")
        self.assertContains(response, "Resumen de compras")
        self.assertContains(response, "Bloqueo prioritario por etapa")
        self.assertContains(response, "Criterios de cierre")
        self.assertContains(response, "Prioridades de atención")
        self.assertContains(response, "Resumen de seguimiento")
        self.assertContains(response, "Control por frente")
        self.assertContains(response, "Dependencias upstream del plan")
        self.assertIn("upstream_dependency_rows", ctx)
        self.assertIn("demand_gate", ctx)
        self.assertIn("branch_priority_rows", ctx)
        self.assertTrue(ctx["branch_priority_rows"])
        self.assertIn("branch_supply_rows", ctx)
        self.assertTrue(ctx["branch_supply_rows"])
        self.assertEqual(
            [row["label"] for row in ctx["upstream_dependency_rows"]],
            ["Plan de producción", "Demanda histórica / forecast", "Maestro de artículos", "Plan y demanda operativa"],
        )
        self.assertContains(response, "Cierre global")
        self.assertIn("erp_command_center", response.context)
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("release_gate_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertIn("erp_governance_rows", response.context)
        self.assertIn("release_gate_completion", response.context)

    def test_plan_scope_context_respects_selected_focus_filters(self):
        SolicitudCompra.objects.create(
            area=f"PLAN_PRODUCCION:{self.plan.id}",
            solicitante="planificador",
            insumo=self.insumo_masa_blank,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("1"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        response = self.client.get(
            reverse("compras:solicitudes"),
            {
                "source": "plan",
                "plan_id": str(self.plan.id),
                "closure_key": "recepciones_aplicadas",
                "handoff_key": "recepcion_cierre",
                "master_class": Insumo.TIPO_MATERIA_PRIMA,
            },
        )
        self.assertEqual(response.status_code, 200)
        ctx = response.context["plan_scope_context"]
        self.assertEqual(ctx["closure_focus"]["label"], "Recepciones aplicadas")
        self.assertEqual(ctx["handoff_focus"]["label"], "Recepción → Cierre")
        self.assertTrue(any(item["is_active"] for item in ctx["closure_checks"] if item["key"] == "recepciones_aplicadas"))
        self.assertTrue(any(item["is_active"] for item in ctx["handoff_checks"] if item["key"] == "recepcion_cierre"))
        self.assertTrue(any(item["is_active"] for item in ctx["master_blocker_class_cards"] if item["class_key"] == Insumo.TIPO_MATERIA_PRIMA))
        self.assertIn("master_blocker_missing_cards", ctx)
        self.assertGreaterEqual(len(ctx["master_blocker_missing_cards"]), 1)
        self.assertTrue(all(row["class_key"] == Insumo.TIPO_MATERIA_PRIMA for row in ctx["master_focus_rows"]))
        self.assertContains(response, "Bloqueos del maestro por clase")
        self.assertContains(response, "Bloqueos del maestro por dato faltante")
        self.assertContains(response, "Enfocar")
        self.assertContains(response, "Bloqueo maestro prioritario")
        self.assertContains(response, "Harina sin categoria")
        self.assertContains(response, "Solicitudes bloqueadas")
        self.assertContains(response, "bloqueados 1")
        self.assertContains(response, "Cierre ")
        self.assertContains(response, "erp-progress")
        self.assertContains(response, "Resolver bloqueos")
        self.assertContains(response, "Solicitudes liberadas")
        self.assertContains(response, "Revisar solicitudes bloqueadas")
        self.assertContains(response, "Corrige bloqueos ERP y completa datos faltantes antes de emitir órdenes.")
        self.assertContains(response, "Criterio de cierre prioritario")
        self.assertContains(response, "El cierre del plan sigue abierto por: recepciones aplicadas.")

    def test_solicitudes_enterprise_board_shows_blocking_articles_detail(self):
        insumo_incompleto = Insumo.objects.create(
            nombre="Chocolate sin proveedor",
            categoria="Cobertura",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=insumo_incompleto,
            proveedor=self.proveedor,
            costo_unitario=Decimal("15"),
            source_hash="cost-choco-sin-proveedor",
        )
        solicitud = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=insumo_incompleto,
            cantidad=Decimal("5"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )

        response = self.client.get(reverse("compras:solicitudes"), {"estatus": "BLOCKED_ERP"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Chocolate sin proveedor")
        self.assertContains(response, "proveedor principal")
        self.assertContains(response, "Asignar proveedor principal")
        self.assertContains(response, "Asigna el proveedor principal")
        self.assertContains(response, solicitud.folio)
        self.assertContains(response, reverse("maestros:insumo_list"))
        self.assertContains(response, f"insumo_id={insumo_incompleto.id}")
        self.assertContains(response, reverse("maestros:insumo_update", args=[insumo_incompleto.id]))
        self.assertContains(response, "Editar artículo")
        self.assertContains(response, "Bloqueos del maestro por clase")
        self.assertContains(response, "Bloqueos del maestro por dato faltante")
        self.assertContains(response, "Detalle del maestro bloqueando solicitudes")
        self.assertContains(response, "Maestro ERP")
        self.assertContains(response, "Incompleto")
        board = response.context["enterprise_board"]
        self.assertTrue(any(item["class_label"] == "Materia prima" for item in board["master_blocker_class_cards"]))
        self.assertTrue(any(item["key"] == "proveedor" for item in board["master_blocker_missing_cards"]))
        self.assertTrue(any(item["class_label"] == "Materia prima" for item in board["master_blocker_detail_rows"]))
        self.assertTrue(
            any(
                item["key"] == "corregir_maestro" and item["query"] == "?workflow_action=corregir_maestro"
                for item in board["next_step_cards"]
            )
        )
        self.assertTrue(
            any(
                item["key"] == "sin_proveedor" and item["query"] == "?blocker_key=sin_proveedor"
                for item in board["blocker_cards"]
            )
        )

    def test_solicitudes_view_can_filter_by_workflow_action_and_blocker_key(self):
        insumo_listo = Insumo.objects.create(
            nombre="Insumo listo filtro workflow solicitud",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
            codigo_point="WF-SOL-001",
        )
        CostoInsumo.objects.create(
            insumo=insumo_listo,
            proveedor=self.proveedor,
            costo_unitario=Decimal("10"),
            source_hash="wf-sol-listo",
        )
        solicitud_lista = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="ana",
            insumo=insumo_listo,
            cantidad=Decimal("2"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        insumo_bloqueado = Insumo.objects.create(
            nombre="Insumo sin proveedor filtro workflow solicitud",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=insumo_bloqueado,
            proveedor=self.proveedor,
            costo_unitario=Decimal("8"),
            source_hash="wf-sol-bloq",
        )
        solicitud_bloqueada = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="ana",
            insumo=insumo_bloqueado,
            cantidad=Decimal("1"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )

        response = self.client.get(reverse("compras:solicitudes"), {"workflow_action": "corregir_maestro"})
        self.assertContains(response, solicitud_bloqueada.folio)
        self.assertNotContains(response, solicitud_lista.folio)

        response = self.client.get(reverse("compras:solicitudes"), {"blocker_key": "sin_proveedor"})
        self.assertContains(response, solicitud_bloqueada.folio)
        self.assertNotContains(response, solicitud_lista.folio)

    def test_solicitudes_view_can_filter_by_closure_and_handoff_key(self):
        insumo_listo = Insumo.objects.create(
            nombre="Insumo listo cierre solicitud",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
            codigo_point="CLS-SOL-001",
        )
        CostoInsumo.objects.create(
            insumo=insumo_listo,
            proveedor=self.proveedor,
            costo_unitario=Decimal("12"),
            source_hash="cls-sol-listo",
        )
        solicitud_borrador = SolicitudCompra.objects.create(
            area=f"PLAN_PRODUCCION:{self.plan.id}",
            solicitante="ana",
            insumo=self.insumo_masa_blank,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("1"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        solicitud_aprobada = SolicitudCompra.objects.create(
            area=f"PLAN_PRODUCCION:{self.plan.id}",
            solicitante="ana",
            insumo=insumo_listo,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("2"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_APROBADA,
        )

        response = self.client.get(
            reverse("compras:solicitudes"),
            {"source": "plan", "plan_id": str(self.plan.id), "closure_key": "solicitudes_liberadas"},
        )
        rendered_solicitudes = list(response.context["solicitudes"])
        self.assertIn(solicitud_borrador, rendered_solicitudes)
        self.assertNotIn(solicitud_aprobada, rendered_solicitudes)
        self.assertEqual(response.context["closure_key_filter"], "solicitudes_liberadas")

        response = self.client.get(
            reverse("compras:solicitudes"),
            {"source": "plan", "plan_id": str(self.plan.id), "handoff_key": "solicitud_orden"},
        )
        rendered_solicitudes = list(response.context["solicitudes"])
        self.assertIn(solicitud_borrador, rendered_solicitudes)
        self.assertIn(solicitud_aprobada, rendered_solicitudes)
        self.assertEqual(response.context["handoff_key_filter"], "solicitud_orden")

    def test_solicitudes_view_can_filter_by_master_blocker(self):
        insumo_mp = Insumo.objects.create(
            nombre="Harina sin proveedor filtro solicitud",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=insumo_mp,
            proveedor=self.proveedor,
            costo_unitario=Decimal("22"),
            source_hash="cost-solicitud-mp-filter",
        )
        solicitud_mp = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=insumo_mp,
            cantidad=Decimal("3"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        insumo_pack = Insumo.objects.create(
            nombre="Etiqueta sin categoria filtro solicitud",
            categoria="",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
            tipo_item=Insumo.TIPO_EMPAQUE,
        )
        CostoInsumo.objects.create(
            insumo=insumo_pack,
            proveedor=self.proveedor,
            costo_unitario=Decimal("3"),
            source_hash="cost-solicitud-pack-filter",
        )
        SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=insumo_pack,
            cantidad=Decimal("1"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )

        response = self.client.get(
            reverse("compras:solicitudes"),
            {"master_class": Insumo.TIPO_MATERIA_PRIMA, "master_missing": "proveedor"},
        )
        self.assertEqual(response.status_code, 200)
        rendered_folios = [item.folio for item in response.context["solicitudes"]]
        self.assertIn(solicitud_mp.folio, rendered_folios)
        self.assertEqual(len(rendered_folios), 1)
        self.assertEqual(response.context["master_class_filter"], Insumo.TIPO_MATERIA_PRIMA)
        self.assertEqual(response.context["master_missing_filter"], "proveedor")
        self.assertContains(response, "Clase maestro")
        self.assertContains(response, "Faltante maestro")

    def test_solicitud_plan_scope_post_forces_area_and_redirects_to_plan(self):
        response = self.client.post(
            reverse("compras:solicitudes"),
            {
                "source": "plan",
                "plan_id": str(self.plan.id),
                "area": "Compras General",
                "solicitante": "planificador",
                "insumo_id": str(self.insumo_masa_explicit.id),
                "cantidad": "4",
                "fecha_requerida": self.fecha_base.isoformat(),
                "estatus": SolicitudCompra.STATUS_BORRADOR,
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("source=plan", response["Location"])
        self.assertIn(f"plan_id={self.plan.id}", response["Location"])

        solicitud = SolicitudCompra.objects.order_by("-id").first()
        self.assertEqual(solicitud.insumo_id, self.insumo_masa_explicit.id)
        self.assertEqual(solicitud.area, f"PLAN_PRODUCCION:{self.plan.id}")
        self.assertEqual(solicitud.solicitante, "planificador")

    def test_solicitud_manual_normaliza_variante_al_canonico(self):
        canonical = Insumo.objects.create(
            nombre="Etiqueta Canonica Compras",
            categoria="Etiquetas",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
            codigo_point="CMP-CAN-01",
        )
        variant = Insumo.objects.create(
            nombre="ETIQUETA CANONICA COMPRAS",
            categoria="Etiquetas",
            unidad_base=self.unidad_kg,
            activo=True,
        )

        response = self.client.post(
            reverse("compras:solicitudes"),
            {
                "area": "Compras",
                "solicitante": "admin",
                "insumo_id": str(variant.id),
                "cantidad": "2",
                "fecha_requerida": "2026-02-10",
                "estatus": SolicitudCompra.STATUS_BORRADOR,
            },
        )
        self.assertEqual(response.status_code, 302)
        created = SolicitudCompra.objects.order_by("-id").first()
        self.assertEqual(created.insumo_id, canonical.id)

    def test_resumen_api_usa_costo_canonico_para_variante_historica(self):
        canonical = Insumo.objects.create(
            nombre="Azucar Canonica Resumen",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
            codigo_point="CMP-SUM-CAN-01",
        )
        variant = Insumo.objects.create(
            nombre="AZUCAR CANONICA RESUMEN",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=canonical,
            proveedor=self.proveedor,
            costo_unitario=Decimal("12.50"),
            source_hash="cost-summary-canonical",
        )
        SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=variant,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("4"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_APROBADA,
        )

        response = self.client.get(
            reverse("compras:solicitudes_resumen_api"),
            {"periodo_tipo": "mes", "periodo_mes": self.periodo_mes, "categoria": "Masa"},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertGreaterEqual(payload["totals"]["presupuesto_estimado_total"], 75.0)

    def test_solicitudes_view_muestra_stock_y_costo_canonico_para_variante_historica(self):
        canonical = Insumo.objects.create(
            nombre="Crema Canonica Solicitudes",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
            codigo_point="CMP-LIST-CAN-01",
        )
        variant = Insumo.objects.create(
            nombre="CREMA CANONICA SOLICITUDES",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=canonical,
            proveedor=self.proveedor,
            costo_unitario=Decimal("9.00"),
            source_hash="cost-list-canonical",
        )
        ExistenciaInsumo.objects.create(
            insumo=canonical,
            stock_actual=Decimal("2"),
            punto_reorden=Decimal("5"),
        )
        ExistenciaInsumo.objects.create(
            insumo=variant,
            stock_actual=Decimal("3"),
            punto_reorden=Decimal("1"),
        )
        solicitud = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=variant,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("2"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )

        response = self.client.get(
            reverse("compras:solicitudes"),
            {"periodo_tipo": "mes", "periodo_mes": self.periodo_mes, "categoria": "Masa"},
        )
        self.assertEqual(response.status_code, 200)
        rendered = next(item for item in response.context["solicitudes"] if item.id == solicitud.id)
        self.assertEqual(rendered.costo_unitario, Decimal("9.00"))
        self.assertEqual(rendered.reabasto_detalle, "Stock 5.000 / Reorden 5.000")

    def test_solicitudes_view_detecta_variante_no_canonica_como_bloqueo_erp(self):
        canonical = Insumo.objects.create(
            nombre="Etiqueta Canonica Bloqueo Compras",
            categoria="Etiquetas",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
            codigo_point="CMP-CAN-BLOCK-01",
        )
        variant = Insumo.objects.create(
            nombre="ETIQUETA CANONICA BLOQUEO COMPRAS",
            categoria="Etiquetas",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=canonical,
            proveedor=self.proveedor,
            costo_unitario=Decimal("4.50"),
            source_hash="cost-canonical-blocker-compras",
        )
        solicitud = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=variant,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("2"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )

        response = self.client.get(reverse("compras:solicitudes"), {"estatus": "BLOCKED_ERP"})
        self.assertEqual(response.status_code, 200)
        rendered = next(item for item in response.context["solicitudes"] if item.id == solicitud.id)
        self.assertTrue(rendered.has_workflow_blockers)
        self.assertIn("variante no canónica", " ".join(rendered.workflow_blockers).lower())
        self.assertEqual(rendered.enterprise_master_summary["status_label"], "Incompleto")
        self.assertContains(response, "Depurar artículo maestro")
        self.assertContains(response, canonical.nombre)

    def test_solicitudes_view_filtra_bloqueo_no_canonico(self):
        canonical = Insumo.objects.create(
            nombre="Caja Canonica Filtro Compras",
            categoria="Empaques",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
            codigo_point="CMP-CAN-FILTER-01",
        )
        variant = Insumo.objects.create(
            nombre="CAJA CANONICA FILTRO COMPRAS",
            categoria="Empaques",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=canonical,
            proveedor=self.proveedor,
            costo_unitario=Decimal("7.00"),
            source_hash="cost-canonical-filter-compras",
        )
        solicitud_variante = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=variant,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("1"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        solicitud_canonica = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=canonical,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("1"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )

        response = self.client.get(reverse("compras:solicitudes"), {"blocker_key": "no_canonico"})
        self.assertEqual(response.status_code, 200)
        rendered_folios = [item.folio for item in response.context["solicitudes"]]
        self.assertIn(solicitud_variante.folio, rendered_folios)
        self.assertNotIn(solicitud_canonica.folio, rendered_folios)

    def test_solicitudes_view_exposes_workflow_summary_and_stage(self):
        response = self.client.get(
            reverse("compras:solicitudes"),
            {"periodo_tipo": "mes", "periodo_mes": self.periodo_mes},
        )
        self.assertEqual(response.status_code, 200)
        workflow_summary = response.context["workflow_summary"]
        cards = {item["label"]: item["count"] for item in workflow_summary["cards"]}
        self.assertEqual(cards["Listas para OC"], 1)
        self.assertIn("gate_cards", workflow_summary)
        self.assertTrue(any(item["key"] == "ready_for_oc" for item in workflow_summary["gate_cards"]))
        self.assertContains(response, "Control Documental")
        self.assertContains(response, "Dependencias del flujo")
        self.assertContains(response, "Solicitudes / BOM")
        self.assertContains(response, "Órdenes documentales")
        self.assertContains(response, "Recepciones / Inventario")
        self.assertContains(response, "Semáforo de etapa")
        self.assertContains(response, "Maestro listo")
        self.assertContains(response, "Lista para compra")
        self.assertContains(response, "Crear orden de compra")
        self.assertIn("trunk_handoff_rows", response.context)
        self.assertEqual(len(response.context["trunk_handoff_rows"]), 3)

    def test_solicitudes_view_exposes_enterprise_board(self):
        insumo_listo = Insumo.objects.create(
            nombre="Insumo listo enterprise",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
            codigo_point="ERP-LISTO-001",
        )
        CostoInsumo.objects.create(
            insumo=insumo_listo,
            proveedor=self.proveedor,
            costo_unitario=Decimal("6.50"),
            source_hash="cost-insumo-listo-enterprise",
        )
        SolicitudCompra.objects.create(
            area="Compras",
            solicitante="ana",
            insumo=insumo_listo,
            cantidad=Decimal("4"),
            fecha_requerida=date(2026, 2, 20),
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        blocked = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="ana",
            insumo=self.insumo_masa_blank,
            cantidad=Decimal("1"),
            fecha_requerida=date(2026, 2, 20),
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        response = self.client.get(
            reverse("compras:solicitudes"),
            {"periodo_tipo": "mes", "periodo_mes": self.periodo_mes},
        )
        self.assertEqual(response.status_code, 200)
        board = response.context["enterprise_board"]
        next_steps = {item["label"]: item["count"] for item in board["next_step_cards"]}
        blockers = {item["label"]: item["count"] for item in board["blocker_cards"]}

        self.assertGreaterEqual(next_steps["Crear OC"], 1)
        self.assertGreaterEqual(next_steps["Corrección ERP"], 1)
        self.assertGreaterEqual(blockers["Maestro incompleto"], 1)
        self.assertIn("master_blocker_class_cards", board)
        self.assertIn("master_blocker_detail_rows", board)
        self.assertContains(response, "Salud operativa ERP")
        self.assertContains(response, "Corregir maestro ERP")
        self.assertContains(response, blocked.folio)

    def test_solicitudes_view_shows_compact_blocker_summary(self):
        insumo_bloqueado = Insumo.objects.create(
            nombre="Insumo bloqueado multiple",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        solicitud = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="ana",
            insumo=insumo_bloqueado,
            cantidad=Decimal("2"),
            fecha_requerida=date(2026, 2, 20),
            estatus=SolicitudCompra.STATUS_EN_REVISION,
        )

        response = self.client.get(
            reverse("compras:solicitudes"),
            {"periodo_tipo": "mes", "periodo_mes": self.periodo_mes},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Bloqueada ERP")
        self.assertContains(response, "Sin costo vigente")
        self.assertContains(response, "más")
        self.assertContains(response, solicitud.folio)
        self.assertContains(response, 'Aprobar</button>')
        self.assertContains(response, 'disabled title="')

    def test_resumen_api_incluye_semaforo_objetivo_por_proveedor_y_categoria(self):
        periodo = PresupuestoCompraPeriodo.objects.create(
            periodo_tipo=PresupuestoCompraPeriodo.TIPO_MES,
            periodo_mes=self.periodo_mes,
            monto_objetivo=Decimal("500"),
            actualizado_por=self.user,
        )
        PresupuestoCompraProveedor.objects.create(
            presupuesto_periodo=periodo,
            proveedor=self.proveedor,
            monto_objetivo=Decimal("20"),
            actualizado_por=self.user,
        )
        PresupuestoCompraCategoria.objects.create(
            presupuesto_periodo=periodo,
            categoria="Masa",
            monto_objetivo=Decimal("25"),
            actualizado_por=self.user,
        )

        response = self.client.get(
            reverse("compras:solicitudes_resumen_api"),
            {
                "periodo_tipo": "mes",
                "periodo_mes": self.periodo_mes,
                "categoria": "Masa",
            },
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()

        self.assertEqual(payload["totals"]["proveedor_objetivo_excedido_count"], 1)
        self.assertEqual(payload["totals"]["categoria_objetivo_excedido_count"], 1)

    def test_resumen_api_usa_proveedor_y_categoria_canonicos_para_variante_historica(self):
        proveedor_canonico = Proveedor.objects.create(nombre="Proveedor Canonico Resumen", activo=True)
        canonical = Insumo.objects.create(
            nombre="Caja Canonica Resumen",
            categoria="Empaques",
            unidad_base=self.unidad_kg,
            proveedor_principal=proveedor_canonico,
            activo=True,
            codigo_point="CMP-CAT-CAN-01",
        )
        variant = Insumo.objects.create(
            nombre="CAJA CANONICA RESUMEN",
            categoria="",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=canonical,
            proveedor=proveedor_canonico,
            costo_unitario=Decimal("8.00"),
            source_hash="cost-category-canonical",
        )
        SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=variant,
            cantidad=Decimal("2"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_APROBADA,
        )

        response = self.client.get(
            reverse("compras:solicitudes_resumen_api"),
            {"periodo_tipo": "mes", "periodo_mes": self.periodo_mes},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()

        proveedores = {row["proveedor"] for row in payload["top_proveedores"]}
        categorias = {row["categoria"] for row in payload["top_categorias"]}
        self.assertIn("Proveedor Canonico Resumen", proveedores)
        self.assertIn("Empaques", categorias)

    def test_export_consolidado_csv_incluye_columna_estado_objetivo(self):
        periodo = PresupuestoCompraPeriodo.objects.create(
            periodo_tipo=PresupuestoCompraPeriodo.TIPO_MES,
            periodo_mes=self.periodo_mes,
            monto_objetivo=Decimal("500"),
            actualizado_por=self.user,
        )
        PresupuestoCompraProveedor.objects.create(
            presupuesto_periodo=periodo,
            proveedor=self.proveedor,
            monto_objetivo=Decimal("20"),
            actualizado_por=self.user,
        )
        PresupuestoCompraCategoria.objects.create(
            presupuesto_periodo=periodo,
            categoria="Masa",
            monto_objetivo=Decimal("25"),
            actualizado_por=self.user,
        )

        response = self.client.get(
            reverse("compras:solicitudes"),
            {
                "periodo_tipo": "mes",
                "periodo_mes": self.periodo_mes,
                "categoria": "Masa",
                "export": "consolidado_csv",
            },
        )
        self.assertEqual(response.status_code, 200)
        body = response.content.decode("utf-8")
        self.assertIn("Estado objetivo", body)
        self.assertIn("Excedido", body)

    def test_consumo_vs_plan_api_retorna_totales(self):
        url = reverse("compras:solicitudes_consumo_vs_plan_api")
        response = self.client.get(
            url,
            {
                "periodo_tipo": "mes",
                "periodo_mes": self.periodo_mes,
                "categoria": "Masa",
                "source": "all",
                "consumo_ref": "all",
            },
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()

        self.assertEqual(payload["filters"]["categoria"], "Masa")
        self.assertEqual(payload["filters"]["consumo_ref"], "all")
        self.assertAlmostEqual(payload["totals"]["plan_qty_total"], 2.0, places=2)
        self.assertAlmostEqual(payload["totals"]["consumo_real_qty_total"], 3.0, places=2)
        self.assertAlmostEqual(payload["totals"]["plan_cost_total"], 20.0, places=2)
        self.assertAlmostEqual(payload["totals"]["consumo_real_cost_total"], 30.0, places=2)
        self.assertAlmostEqual(payload["totals"]["variacion_cost_total"], 10.0, places=2)
        self.assertIsNotNone(payload["totals"]["cobertura_pct"])
        self.assertEqual(payload["totals"]["sin_costo_count"], 0)
        self.assertEqual(payload["totals"]["semaforo_rojo_count"], 1)
        self.assertEqual(payload["rows"][0]["semaforo"], "ROJO")
        self.assertFalse(payload["rows"][0]["sin_costo"])

    def test_consumo_vs_plan_api_soporta_offset_y_sort(self):
        insumo_extra = Insumo.objects.create(
            nombre="Masa extra API",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=insumo_extra,
            proveedor=self.proveedor,
            costo_unitario=Decimal("4"),
            source_hash="cost-masa-extra-1",
        )
        receta_extra = Receta.objects.create(
            nombre="Receta extra consumo API",
            hash_contenido="hash-consumo-api-extra",
        )
        LineaReceta.objects.create(
            receta=receta_extra,
            posicion=1,
            insumo=insumo_extra,
            insumo_texto="Masa extra API",
            cantidad=Decimal("1"),
            unidad=self.unidad_kg,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("4"),
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100,
            match_status=LineaReceta.STATUS_AUTO,
        )
        plan_extra = PlanProduccion.objects.create(
            nombre="Plan extra consumo",
            fecha_produccion=self.fecha_base,
        )
        PlanProduccionItem.objects.create(
            plan=plan_extra,
            receta=receta_extra,
            cantidad=Decimal("1"),
        )

        url = reverse("compras:solicitudes_consumo_vs_plan_api")
        response = self.client.get(
            url,
            {
                "periodo_tipo": "mes",
                "periodo_mes": self.periodo_mes,
                "categoria": "Masa",
                "source": "all",
                "consumo_ref": "all",
                "limit": 1,
                "offset": 1,
                "sort_by": "variacion_cost_abs",
                "sort_dir": "desc",
            },
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["filters"]["limit"], 1)
        self.assertEqual(payload["filters"]["offset"], 1)
        self.assertEqual(payload["filters"]["sort_by"], "variacion_cost_abs")
        self.assertEqual(payload["filters"]["sort_dir"], "desc")
        self.assertEqual(payload["meta"]["rows_total"], 2)
        self.assertEqual(payload["meta"]["rows_returned"], 1)
        self.assertEqual(payload["rows"][0]["insumo"], "Masa extra API")

    def test_consumo_vs_plan_api_sort_invalido(self):
        url = reverse("compras:solicitudes_consumo_vs_plan_api")
        bad_sort = self.client.get(url, {"sort_by": "ruido"})
        self.assertEqual(bad_sort.status_code, 400)
        self.assertIn("sort_by", bad_sort.json()["detail"].lower())

        bad_dir = self.client.get(url, {"sort_dir": "up"})
        self.assertEqual(bad_dir.status_code, 400)
        self.assertIn("sort_dir", bad_dir.json()["detail"].lower())

    def test_consumo_vs_plan_api_filtra_movimientos_solo_con_referencia_plan(self):
        MovimientoInventario.objects.create(
            fecha=timezone.make_aware(datetime(2026, 2, 10, 12, 0, 0)),
            tipo=MovimientoInventario.TIPO_CONSUMO,
            insumo=self.insumo_masa_blank,
            cantidad=Decimal("2"),
            referencia="SALIDA_MANUAL",
        )
        url = reverse("compras:solicitudes_consumo_vs_plan_api")
        payload_all = self.client.get(
            url,
            {
                "periodo_tipo": "mes",
                "periodo_mes": self.periodo_mes,
                "categoria": "Masa",
                "source": "all",
                "consumo_ref": "all",
            },
        ).json()
        payload_plan_ref = self.client.get(
            url,
            {
                "periodo_tipo": "mes",
                "periodo_mes": self.periodo_mes,
                "categoria": "Masa",
                "source": "all",
                "consumo_ref": "plan_ref",
            },
        ).json()

        self.assertAlmostEqual(payload_all["totals"]["consumo_real_qty_total"], 5.0, places=2)
        self.assertAlmostEqual(payload_plan_ref["totals"]["consumo_real_qty_total"], 3.0, places=2)
        self.assertEqual(payload_plan_ref["filters"]["consumo_ref"], "plan_ref")

    def test_consumo_vs_plan_marca_alerta_sin_costo_unitario(self):
        insumo_sin_costo = Insumo.objects.create(
            nombre="Insumo interno sin costo",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        receta_sin_costo = Receta.objects.create(
            nombre="Receta sin costo",
            hash_contenido="test-hash-plan-002",
        )
        LineaReceta.objects.create(
            receta=receta_sin_costo,
            posicion=1,
            insumo=insumo_sin_costo,
            insumo_texto="Insumo interno sin costo",
            cantidad=Decimal("1"),
            unidad=self.unidad_kg,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("0"),
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100,
            match_status=LineaReceta.STATUS_AUTO,
        )
        plan_sin_costo = PlanProduccion.objects.create(
            nombre="Plan Sin Costo",
            fecha_produccion=self.fecha_base,
        )
        PlanProduccionItem.objects.create(
            plan=plan_sin_costo,
            receta=receta_sin_costo,
            cantidad=Decimal("1"),
        )

        url = reverse("compras:solicitudes_consumo_vs_plan_api")
        payload = self.client.get(
            url,
            {
                "periodo_tipo": "mes",
                "periodo_mes": self.periodo_mes,
                "categoria": "Masa",
                "source": "all",
                "consumo_ref": "all",
            },
        ).json()

        self.assertGreaterEqual(payload["totals"]["sin_costo_count"], 1)
        row = next(r for r in payload["rows"] if r["insumo"] == "Insumo interno sin costo")
        self.assertTrue(row["sin_costo"])
        self.assertEqual(row["alerta"], "Sin costo unitario")

    def test_nueva_solicitud_usa_sugerencia_con_seguridad_leadtime_y_transito(self):
        from inventario.models import ExistenciaInsumo

        ExistenciaInsumo.objects.create(
            insumo=self.insumo_masa_blank,
            stock_actual=Decimal("2"),
            punto_reorden=Decimal("5"),
            stock_minimo=Decimal("2"),
            dias_llegada_pedido=2,
            consumo_diario_promedio=Decimal("1.5"),
        )

        response = self.client.get(reverse("compras:solicitudes"))
        self.assertEqual(response.status_code, 200)
        options = response.context["insumo_options"]
        row = next(o for o in options if o["id"] == self.insumo_masa_blank.id)

        self.assertEqual(row["stock_actual"], Decimal("2"))
        self.assertEqual(row["stock_seguridad"], Decimal("2"))
        self.assertEqual(row["demanda_lead_time"], Decimal("3.0"))
        self.assertEqual(row["en_transito"], Decimal("2"))
        self.assertEqual(row["recomendado"], Decimal("1.0"))
        self.assertEqual(row["enterprise_status"], "Incompleto")
        self.assertIn("código Point", row["enterprise_missing"])
        self.assertTrue(row["is_operational_blocker"])
        self.assertEqual(row["operational_blocker_label"], "Bloquea compras")
        self.assertContains(response, "estado ERP")
        self.assertContains(response, "Bloquea compras")

    def test_solicitudes_view_puede_filtrar_bloqueadas_erp(self):
        SolicitudCompra.objects.create(
            area="Compras",
            solicitante="ana",
            insumo=self.insumo_masa_blank,
            cantidad=Decimal("1"),
            fecha_requerida=date(2026, 2, 20),
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        response = self.client.get(
            reverse("compras:solicitudes"),
            {"periodo_tipo": "mes", "periodo_mes": self.periodo_mes, "estatus": "BLOCKED_ERP"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Bloqueadas ERP")
        self.assertTrue(all(getattr(s, "has_workflow_blockers", False) for s in response.context["solicitudes"]))

    def test_nueva_solicitud_insumo_options_no_limita_a_200(self):
        for idx in range(0, 230):
            Insumo.objects.create(
                nombre=f"Insumo extra {idx:03d}",
                categoria="Masa",
                unidad_base=self.unidad_kg,
                proveedor_principal=self.proveedor,
                activo=True,
            )

        response = self.client.get(reverse("compras:solicitudes"))
        self.assertEqual(response.status_code, 200)
        options = response.context["insumo_options"]
        # Debe incluir más de 200 activos para evitar truncamiento operativo.
        self.assertGreater(len(options), 200)

    def test_nueva_solicitud_insumo_options_usa_canonicos_y_oculta_variantes(self):
        canonical = Insumo.objects.create(
            nombre="Azucar Canonica Compras",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
            codigo_point="PT-001",
        )
        variant = Insumo.objects.create(
            nombre="Azucar Canonica Compras ",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=canonical,
            proveedor=self.proveedor,
            costo_unitario=Decimal("9.50"),
            source_hash="cost-canonical-compras",
        )
        CostoInsumo.objects.create(
            insumo=variant,
            proveedor=self.proveedor,
            costo_unitario=Decimal("8.00"),
            source_hash="cost-variant-compras",
        )

        response = self.client.get(reverse("compras:solicitudes"))
        self.assertEqual(response.status_code, 200)
        options = response.context["insumo_options"]

        ids = {row["id"] for row in options}
        self.assertIn(canonical.id, ids)
        self.assertNotIn(variant.id, ids)

        row = next(item for item in options if item["id"] == canonical.id)
        self.assertEqual(row["canonical_variant_count"], 2)
        self.assertEqual(row["enterprise_status"], "Lista para operar")
        self.assertEqual(row["enterprise_missing"], [])
        self.assertFalse(row["is_operational_blocker"])

    def test_nueva_solicitud_agrega_stock_y_transito_de_variantes_al_canonico(self):
        canonical = Insumo.objects.create(
            nombre="Chocolate Canonico Compras",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
            codigo_point="PT-CHOCO",
        )
        variant = Insumo.objects.create(
            nombre="Chocolate Canonico Compras ",
            categoria="Masa",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        ExistenciaInsumo.objects.create(
            insumo=canonical,
            stock_actual=Decimal("2"),
            punto_reorden=Decimal("5"),
            stock_minimo=Decimal("2"),
            dias_llegada_pedido=2,
            consumo_diario_promedio=Decimal("1.5"),
        )
        ExistenciaInsumo.objects.create(
            insumo=variant,
            stock_actual=Decimal("3"),
            punto_reorden=Decimal("0"),
            stock_minimo=Decimal("0"),
            dias_llegada_pedido=0,
            consumo_diario_promedio=Decimal("0"),
        )
        solicitud_variant = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=variant,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("1"),
            fecha_requerida=self.fecha_base,
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        OrdenCompra.objects.create(
            solicitud=solicitud_variant,
            proveedor=self.proveedor,
            fecha_emision=self.fecha_base,
            monto_estimado=Decimal("10"),
            estatus=OrdenCompra.STATUS_ENVIADA,
        )

        response = self.client.get(reverse("compras:solicitudes"))
        self.assertEqual(response.status_code, 200)
        row = next(item for item in response.context["insumo_options"] if item["id"] == canonical.id)

        self.assertEqual(row["stock_actual"], Decimal("5"))
        self.assertEqual(row["en_transito"], Decimal("1"))
        self.assertEqual(row["stock_seguridad"], Decimal("2"))
        self.assertEqual(row["demanda_lead_time"], Decimal("3.0"))
        self.assertEqual(row["recomendado"], Decimal("0"))
        self.assertEqual(row["enterprise_status"], "Lista para operar")
        self.assertFalse(row["is_operational_blocker"])


class ComprasSolicitudesImportPreviewTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_import",
            email="admin_import@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

        self.proveedor = Proveedor.objects.create(nombre="Proveedor Import", activo=True)
        self.unidad = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.insumo_harina = Insumo.objects.create(
            nombre="Harina Import",
            categoria="Masa",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        self.insumo_azucar = Insumo.objects.create(
            nombre="Azucar Import",
            categoria="Masa",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor,
            activo=True,
        )

    def test_import_preview_confirma_edicion_y_descarte(self):
        csv_content = (
            "insumo,cantidad,area,solicitante,fecha_requerida,estatus\n"
            "Harina Import,2,Compras,ana,2026-02-20,BORRADOR\n"
            "Azucar con typo,3,Compras,luis,2026-02-21,BORRADOR\n"
            "Harina Import,1,Compras,maria,2026-02-22,BORRADOR\n"
        )
        archivo = SimpleUploadedFile("solicitudes.csv", csv_content.encode("utf-8"), content_type="text/csv")

        with patch(
            "compras.views.match_insumo",
            side_effect=[
                (self.insumo_harina, 100.0, "exact"),
                (self.insumo_azucar, 60.0, "fuzzy"),
                (self.insumo_harina, 100.0, "exact"),
            ],
        ):
            response = self.client.post(
                reverse("compras:solicitudes_importar"),
                {
                    "archivo": archivo,
                    "periodo_tipo": "mes",
                    "periodo_mes": "2026-02",
                    "area": "Compras",
                    "solicitante": "admin_import",
                    "estatus": SolicitudCompra.STATUS_BORRADOR,
                    "score_min": "90",
                    "evitar_duplicados": "on",
                },
            )

        self.assertEqual(response.status_code, 302)
        preview_payload = self.client.session.get("compras_solicitudes_import_preview")
        self.assertIsNotNone(preview_payload)
        self.assertEqual(SolicitudCompra.objects.count(), 0)
        rows = preview_payload["rows"]
        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[1]["insumo_id"], "")

        confirmar_data = {}
        for row in rows:
            row_id = row["row_id"]
            confirmar_data[f"row_{row_id}_insumo_id"] = row["insumo_id"]
            confirmar_data[f"row_{row_id}_cantidad"] = row["cantidad"]
            confirmar_data[f"row_{row_id}_fecha_requerida"] = row["fecha_requerida"]
            confirmar_data[f"row_{row_id}_area"] = row["area"]
            confirmar_data[f"row_{row_id}_solicitante"] = row["solicitante"]
            confirmar_data[f"row_{row_id}_proveedor_id"] = row["proveedor_id"]
            confirmar_data[f"row_{row_id}_estatus"] = row["estatus"]

        confirmar_data[f"row_{rows[0]['row_id']}_include"] = "on"
        confirmar_data[f"row_{rows[1]['row_id']}_include"] = "on"
        confirmar_data[f"row_{rows[1]['row_id']}_insumo_id"] = str(self.insumo_azucar.id)

        response_confirm = self.client.post(
            reverse("compras:solicitudes_importar_confirmar"),
            confirmar_data,
        )
        self.assertEqual(response_confirm.status_code, 302)

        solicitudes = list(SolicitudCompra.objects.order_by("id"))
        self.assertEqual(len(solicitudes), 2)
        self.assertEqual({s.insumo_id for s in solicitudes}, {self.insumo_harina.id, self.insumo_azucar.id})
        self.assertIsNone(self.client.session.get("compras_solicitudes_import_preview"))

    def test_descargar_plantilla_solicitudes_csv(self):
        response = self.client.get(
            reverse("compras:solicitudes_importar_plantilla"),
            {"format": "csv"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        body = response.content.decode("utf-8")
        self.assertIn("insumo,cantidad,proveedor,fecha_requerida,area,solicitante,estatus", body)

    def test_descargar_plantilla_solicitudes_xlsx(self):
        response = self.client.get(
            reverse("compras:solicitudes_importar_plantilla"),
            {"format": "xlsx"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", response["Content-Type"])
        wb = load_workbook(BytesIO(response.content), data_only=True)
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, 8)]
        self.assertEqual(
            headers,
            ["insumo", "cantidad", "proveedor", "fecha_requerida", "area", "solicitante", "estatus"],
        )

    def test_eliminar_solicitud_permitida_si_no_tiene_oc_activa(self):
        solicitud = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin_import",
            insumo=self.insumo_harina,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("1"),
            fecha_requerida=date(2026, 2, 20),
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )

        response = self.client.post(
            reverse("compras:solicitud_eliminar", args=[solicitud.id]),
            {"return_query": "source=manual"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(SolicitudCompra.objects.filter(id=solicitud.id).exists())

    def test_eliminar_solicitud_bloqueada_si_tiene_oc_activa(self):
        solicitud = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin_import",
            insumo=self.insumo_harina,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("1"),
            fecha_requerida=date(2026, 2, 20),
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        OrdenCompra.objects.create(
            solicitud=solicitud,
            proveedor=self.proveedor,
            fecha_emision=date(2026, 2, 20),
            monto_estimado=Decimal("10"),
            estatus=OrdenCompra.STATUS_ENVIADA,
        )

        response = self.client.post(
            reverse("compras:solicitud_eliminar", args=[solicitud.id]),
            {"return_query": "source=manual"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(SolicitudCompra.objects.filter(id=solicitud.id).exists())

    def test_import_preview_summary_counts_in_context(self):
        session = self.client.session
        session["compras_solicitudes_import_preview"] = {
            "file_name": "solicitudes_test.csv",
            "generated_at": "2026-02-20 09:10",
            "score_min": 90,
            "evitar_duplicados": True,
            "rows": [
                {
                    "row_id": "2",
                    "include": True,
                    "duplicate": False,
                    "insumo_id": str(self.insumo_harina.id),
                    "cantidad": "2.000",
                    "costo_unitario": "10.00",
                    "presupuesto_estimado": "20.00",
                    "notes": "",
                },
                {
                    "row_id": "3",
                    "include": False,
                    "duplicate": True,
                    "insumo_id": "",
                    "cantidad": "0",
                    "notes": "Sin match de insumo.",
                },
            ],
        }
        session.save()

        response = self.client.get(reverse("compras:solicitudes"))
        self.assertEqual(response.status_code, 200)
        preview = response.context["import_preview"]
        self.assertEqual(preview["count"], 2)
        self.assertEqual(preview["ready_count"], 1)
        self.assertEqual(preview["excluded_count"], 1)
        self.assertEqual(preview["issues_count"], 1)
        self.assertEqual(preview["duplicates_count"], 1)
        self.assertEqual(preview["without_match_count"], 1)
        self.assertEqual(preview["invalid_qty_count"], 1)
        self.assertEqual(preview["ready_qty_total"], Decimal("2.000"))
        self.assertEqual(preview["ready_budget_total"], Decimal("20.00"))
        self.assertEqual(preview["file_name"], "solicitudes_test.csv")
        self.assertEqual(preview["generated_at"], "2026-02-20 09:10")

    def test_export_import_preview_csv(self):
        session = self.client.session
        session["compras_solicitudes_import_preview"] = {
            "rows": [
                {
                    "row_id": "2",
                    "source_row": 2,
                    "include": True,
                    "insumo_origen": "Harina Import",
                    "insumo_sugerencia": "Harina Import",
                    "insumo_id": str(self.insumo_harina.id),
                    "cantidad": "2.000",
                    "area": "Compras",
                    "solicitante": "ana",
                    "fecha_requerida": "2026-02-20",
                    "estatus": SolicitudCompra.STATUS_BORRADOR,
                    "proveedor_id": str(self.proveedor.id),
                    "score": "100.0",
                    "metodo": "EXACT",
                    "costo_unitario": "10.00",
                    "presupuesto_estimado": "20.00",
                    "duplicate": False,
                    "notes": "",
                }
            ]
        }
        session.save()

        response = self.client.get(reverse("compras:solicitudes"), {"export": "import_preview_csv"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        body = response.content.decode("utf-8")
        self.assertIn("row_id,source_row,include,insumo_origen", body)
        self.assertIn("costo_unitario,presupuesto_estimado", body)
        self.assertIn("Harina Import", body)

    def test_export_import_preview_xlsx(self):
        session = self.client.session
        session["compras_solicitudes_import_preview"] = {
            "rows": [
                {
                    "row_id": "2",
                    "source_row": 2,
                    "include": True,
                    "insumo_origen": "Harina Import",
                    "insumo_sugerencia": "Harina Import",
                    "insumo_id": str(self.insumo_harina.id),
                    "cantidad": "2.000",
                    "area": "Compras",
                    "solicitante": "ana",
                    "fecha_requerida": "2026-02-20",
                    "estatus": SolicitudCompra.STATUS_BORRADOR,
                    "proveedor_id": str(self.proveedor.id),
                    "score": "100.0",
                    "metodo": "EXACT",
                    "costo_unitario": "10.00",
                    "presupuesto_estimado": "20.00",
                    "duplicate": False,
                    "notes": "",
                }
            ]
        }
        session.save()

        response = self.client.get(reverse("compras:solicitudes"), {"export": "import_preview_xlsx"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", response["Content-Type"])

        wb = load_workbook(BytesIO(response.content), data_only=True)
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, 19)]
        self.assertEqual(
            headers,
            [
                "row_id",
                "source_row",
                "include",
                "insumo_origen",
                "insumo_sugerencia",
                "insumo_id",
                "cantidad",
                "area",
                "solicitante",
                "fecha_requerida",
                "estatus",
                "proveedor_id",
                "score",
                "metodo",
                "costo_unitario",
                "presupuesto_estimado",
                "duplicate",
                "notes",
            ],
        )
        self.assertEqual(ws.cell(row=2, column=4).value, "Harina Import")

    def test_import_preview_detecta_duplicado_en_lote(self):
        SolicitudCompra.objects.create(
            area="Compras",
            solicitante="ana",
            insumo=self.insumo_harina,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("2"),
            fecha_requerida=date(2026, 2, 20),
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        csv_content = (
            "insumo,cantidad,area,solicitante,fecha_requerida,estatus\n"
            "Harina Import,2,Compras,ana,2026-02-20,BORRADOR\n"
        )
        archivo = SimpleUploadedFile("solicitudes_dup.csv", csv_content.encode("utf-8"), content_type="text/csv")

        with patch(
            "compras.views.match_insumo",
            return_value=(self.insumo_harina, 100.0, "exact"),
        ):
            response = self.client.post(
                reverse("compras:solicitudes_importar"),
                {
                    "archivo": archivo,
                    "periodo_tipo": "mes",
                    "periodo_mes": "2026-02",
                    "area": "Compras",
                    "solicitante": "ana",
                    "estatus": SolicitudCompra.STATUS_BORRADOR,
                    "score_min": "90",
                    "evitar_duplicados": "on",
                },
            )

        self.assertEqual(response.status_code, 302)
        preview_payload = self.client.session.get("compras_solicitudes_import_preview")
        self.assertIsNotNone(preview_payload)
        self.assertEqual(len(preview_payload["rows"]), 1)
        row = preview_payload["rows"][0]
        self.assertTrue(row["duplicate"])
        self.assertIn("Posible duplicado", row["notes"])

    def test_confirm_import_evita_duplicados_dentro_del_mismo_archivo(self):
        session = self.client.session
        session["compras_solicitudes_import_preview"] = {
            "periodo_tipo": "mes",
            "periodo_mes": "2026-02",
            "evitar_duplicados": True,
            "rows": [
                {
                    "row_id": "2",
                    "source_row": 2,
                    "insumo_origen": "Harina Import",
                    "insumo_sugerencia": "Harina Import",
                    "insumo_id": str(self.insumo_harina.id),
                    "cantidad": "2.000",
                    "area": "Compras",
                    "solicitante": "ana",
                    "fecha_requerida": "2026-02-20",
                    "estatus": SolicitudCompra.STATUS_BORRADOR,
                    "proveedor_id": str(self.proveedor.id),
                    "score": "100.0",
                    "metodo": "EXACT",
                    "duplicate": False,
                    "notes": "",
                    "include": True,
                },
                {
                    "row_id": "3",
                    "source_row": 3,
                    "insumo_origen": "Harina Import",
                    "insumo_sugerencia": "Harina Import",
                    "insumo_id": str(self.insumo_harina.id),
                    "cantidad": "1.000",
                    "area": "Compras",
                    "solicitante": "ana",
                    "fecha_requerida": "2026-02-20",
                    "estatus": SolicitudCompra.STATUS_BORRADOR,
                    "proveedor_id": str(self.proveedor.id),
                    "score": "100.0",
                    "metodo": "EXACT",
                    "duplicate": False,
                    "notes": "",
                    "include": True,
                },
            ],
        }
        session.save()

        response = self.client.post(
            reverse("compras:solicitudes_importar_confirmar"),
            {
                "row_2_include": "on",
                "row_2_insumo_id": str(self.insumo_harina.id),
                "row_2_cantidad": "2.000",
                "row_2_fecha_requerida": "2026-02-20",
                "row_2_area": "Compras",
                "row_2_solicitante": "ana",
                "row_2_proveedor_id": str(self.proveedor.id),
                "row_2_estatus": SolicitudCompra.STATUS_BORRADOR,
                "row_3_include": "on",
                "row_3_insumo_id": str(self.insumo_harina.id),
                "row_3_cantidad": "1.000",
                "row_3_fecha_requerida": "2026-02-20",
                "row_3_area": "Compras",
                "row_3_solicitante": "ana",
                "row_3_proveedor_id": str(self.proveedor.id),
                "row_3_estatus": SolicitudCompra.STATUS_BORRADOR,
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(SolicitudCompra.objects.count(), 1)


class ComprasOrdenesRecepcionesFiltersTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_filtros_oc",
            email="admin_filtros_oc@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

        self.unidad = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.proveedor_a = Proveedor.objects.create(nombre="Proveedor A", activo=True)
        self.proveedor_b = Proveedor.objects.create(nombre="Proveedor B", activo=True)
        self.insumo = Insumo.objects.create(
            nombre="Harina OC",
            categoria="Masa",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor_a,
            activo=True,
        )
        self.solicitud = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=self.insumo,
            proveedor_sugerido=self.proveedor_a,
            cantidad=Decimal("2"),
            fecha_requerida=date(2026, 2, 20),
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        self.orden_enviada = OrdenCompra.objects.create(
            solicitud=self.solicitud,
            referencia="OC ENVIADA TEST",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 20),
            monto_estimado=Decimal("100"),
            estatus=OrdenCompra.STATUS_ENVIADA,
        )
        self.orden_cerrada = OrdenCompra.objects.create(
            solicitud=self.solicitud,
            referencia="OC CERRADA TEST",
            proveedor=self.proveedor_b,
            fecha_emision=date(2026, 1, 15),
            monto_estimado=Decimal("50"),
            estatus=OrdenCompra.STATUS_CERRADA,
        )
        self.recepcion_pendiente = RecepcionCompra.objects.create(
            orden=self.orden_enviada,
            fecha_recepcion=date(2026, 2, 21),
            conformidad_pct=Decimal("95"),
            estatus=RecepcionCompra.STATUS_PENDIENTE,
            observaciones="Recepcion parcial",
        )
        self.recepcion_cerrada = RecepcionCompra.objects.create(
            orden=self.orden_cerrada,
            fecha_recepcion=date(2026, 1, 18),
            conformidad_pct=Decimal("100"),
            estatus=RecepcionCompra.STATUS_CERRADA,
            observaciones="Recepcion ok",
        )
        self.plan = PlanProduccion.objects.create(
            nombre="Plan Documental Compras",
            fecha_produccion=date(2026, 2, 21),
        )

    def test_ordenes_filter_by_status_and_month(self):
        response = self.client.get(
            reverse("compras:ordenes"),
            {"estatus": OrdenCompra.STATUS_ENVIADA, "mes": "2026-02"},
        )
        self.assertEqual(response.status_code, 200)
        ordenes = list(response.context["ordenes"])
        self.assertEqual(len(ordenes), 1)
        self.assertEqual(ordenes[0].id, self.orden_enviada.id)

    def test_ordenes_export_csv_respects_filters(self):
        self.orden_enviada.referencia = f"PLAN_PRODUCCION:{self.plan.id}"
        self.orden_enviada.save(update_fields=["referencia"])
        response = self.client.get(
            reverse("compras:ordenes"),
            {"estatus": OrdenCompra.STATUS_ENVIADA, "export": "csv"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        body = response.content.decode("utf-8")
        self.assertIn(self.orden_enviada.folio, body)
        self.assertNotIn(self.orden_cerrada.folio, body)
        self.assertIn("origen", body.splitlines()[0])
        self.assertIn("plan_origen", body.splitlines()[0])
        self.assertIn("plan", body)

    def test_recepciones_filter_by_status_and_month(self):
        response = self.client.get(
            reverse("compras:recepciones"),
            {"estatus": RecepcionCompra.STATUS_PENDIENTE, "mes": "2026-02"},
        )
        self.assertEqual(response.status_code, 200)
        recepciones = list(response.context["recepciones"])
        self.assertEqual(len(recepciones), 1)
        self.assertEqual(recepciones[0].id, self.recepcion_pendiente.id)

    def test_recepciones_filter_by_plan_reference_query(self):
        self.orden_enviada.referencia = "PLAN_PRODUCCION:77"
        self.orden_enviada.save(update_fields=["referencia"])
        response = self.client.get(
            reverse("compras:recepciones"),
            {"q": "PLAN_PRODUCCION:77"},
        )
        self.assertEqual(response.status_code, 200)
        recepciones = list(response.context["recepciones"])
        self.assertEqual(len(recepciones), 1)
        self.assertEqual(recepciones[0].id, self.recepcion_pendiente.id)

    def test_ordenes_view_exposes_plan_scope_context_from_query(self):
        self.orden_enviada.referencia = f"PLAN_PRODUCCION:{self.plan.id}"
        self.orden_enviada.monto_estimado = Decimal("0")
        self.orden_enviada.save(update_fields=["referencia", "monto_estimado"])
        self.solicitud.area = f"PLAN_PRODUCCION:{self.plan.id}"
        self.solicitud.save(update_fields=["area"])
        other_solicitud = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="otro",
            insumo=self.insumo,
            proveedor_sugerido=self.proveedor_a,
            cantidad=Decimal("1"),
            fecha_requerida=date(2026, 2, 22),
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        response = self.client.get(
            reverse("compras:ordenes"),
            {"q": f"PLAN_PRODUCCION:{self.plan.id}"},
        )
        self.assertEqual(response.status_code, 200)
        ctx = response.context["plan_scope_context"]
        self.assertIsNotNone(ctx)
        self.assertIn("master_demand_gate", ctx)
        self.assertIn("branch_priority_rows", ctx)
        self.assertIn("branch_supply_rows", ctx)
        self.assertEqual(ctx["plan_id"], self.plan.id)
        self.assertEqual(ctx["summary_label"], "Con bloqueos")
        self.assertEqual(ctx["stage_label"], "Órdenes bloqueadas ERP")
        self.assertIn("estatus=BLOCKED_ERP", ctx["next_action"]["url"])
        self.assertEqual(ctx["pipeline_steps"][1]["label"], "Órdenes")
        self.assertEqual(ctx["pipeline_steps"][1]["semaphore_label"], "Rojo")
        self.assertEqual(ctx["pipeline_steps"][1]["blocked"], 1)
        self.assertTrue(ctx["pipeline_steps"][1]["is_active"])
        self.assertEqual(ctx["pipeline_steps"][1]["action_label"], "Corregir órdenes")
        self.assertTrue(ctx["master_focus_rows"])
        focus_row = ctx["master_focus_rows"][0]
        self.assertIn(f"insumo_id={self.insumo.id}", focus_row["action_url"])
        self.assertEqual(focus_row["edit_url"], reverse("maestros:insumo_update", args=[self.insumo.id]))
        self.assertContains(response, "Editar artículo")
        self.assertIn("Completa proveedor, monto y datos documentales", ctx["pipeline_steps"][1]["action_detail"])
        self.assertEqual(ctx["stage_focus"]["label"], "Órdenes")
        self.assertEqual(len(ctx["stage_focus"]["rows"]), 1)
        self.assertGreaterEqual(ctx["pipeline_steps"][1]["progress_pct"], 0)
        self.assertEqual(ctx["closure_checks"][1]["label"], "Órdenes sin bloqueo")
        self.assertEqual(ctx["closure_checks"][1]["action_label"], "Corregir órdenes bloqueadas")
        self.assertIn("closure_key=ordenes_sin_bloqueo", ctx["closure_checks"][1]["action_url"])
        self.assertIn("Completa proveedor, monto estimado", ctx["closure_checks"][1]["action_detail"])
        self.assertEqual(ctx["closure_focus"]["label"], "Órdenes sin bloqueo")
        self.assertGreaterEqual(len(ctx["closure_focus_rows"]), 1)
        self.assertEqual(ctx["closure_focus_rows"][0]["scope"], "Orden")
        self.assertEqual(ctx["handoff_checks"][1]["label"], "Orden → Recepción")
        self.assertEqual(ctx["handoff_checks"][1]["action_label"], "Corregir órdenes bloqueadas")
        self.assertIn("handoff_key=orden_recepcion", ctx["handoff_checks"][1]["action_url"])
        self.assertEqual(ctx["handoff_focus"]["label"], "Orden → Recepción")
        self.assertGreaterEqual(len(ctx["handoff_focus_rows"]), 1)
        self.assertEqual(ctx["handoff_focus_rows"][0]["scope"], "Orden")
        self.assertIn("master_focus", ctx)
        self.assertEqual(ctx["master_focus"]["class_label"], "Materia prima")
        self.assertGreaterEqual(len(ctx["master_focus_rows"]), 1)
        self.assertContains(response, self.plan.nombre)
        self.assertContains(response, f"source=plan&amp;plan_id={self.plan.id}")
        self.assertContains(response, "bloqueados 1")
        self.assertContains(response, "Cierre ")
        self.assertContains(response, "erp-progress")
        self.assertContains(response, "Bloqueo prioritario por etapa")
        self.assertContains(response, "Corregir órdenes")
        self.assertEqual(list(response.context["solicitudes"]), [self.solicitud])
        self.assertNotIn(other_solicitud, list(response.context["solicitudes"]))
        orden = list(response.context["ordenes"])[0]
        self.assertEqual(orden.source_tipo, "plan")
        self.assertEqual(orden.source_plan_id, self.plan.id)
        self.assertEqual(orden.source_plan_nombre, self.plan.nombre)
        self.assertContains(response, self.plan.nombre)
        self.assertContains(response, "Bloqueos del plan")
        self.assertContains(response, "Órdenes bloqueadas")
        self.assertContains(response, "Etapa documental actual")
        self.assertContains(response, "Bloqueo maestro prioritario")
        self.assertContains(response, "Dependencias upstream del plan")
        self.assertIn("upstream_dependency_rows", ctx)
        self.assertEqual([row["label"] for row in ctx["upstream_dependency_rows"]], ["Plan de producción", "Maestro de artículos", "Solicitudes liberadas"])
        self.assertContains(response, "Órdenes sin bloqueo")
        self.assertContains(response, "Corregir órdenes bloqueadas")
        self.assertContains(response, "Completa proveedor, monto estimado y datos documentales para habilitar recepción.")
        self.assertContains(response, "El cierre del plan sigue abierto por: órdenes sin bloqueo.")
        self.assertContains(response, "<th>Ámbito</th>", html=False)

    def test_recepciones_view_exposes_plan_scope_context_from_query(self):
        self.orden_enviada.referencia = f"PLAN_PRODUCCION:{self.plan.id}"
        self.orden_enviada.save(update_fields=["referencia"])
        self.recepcion_pendiente.conformidad_pct = Decimal("0")
        self.recepcion_pendiente.save(update_fields=["conformidad_pct"])
        other_orden = OrdenCompra.objects.create(
            solicitud=self.solicitud,
            referencia="COMPRA GENERAL",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 22),
            monto_estimado=Decimal("80"),
            estatus=OrdenCompra.STATUS_CONFIRMADA,
        )
        response = self.client.get(
            reverse("compras:recepciones"),
            {"q": f"PLAN_PRODUCCION:{self.plan.id}"},
        )
        self.assertEqual(response.status_code, 200)
        ctx = response.context["plan_scope_context"]
        self.assertIsNotNone(ctx)
        self.assertIn("master_demand_gate", ctx)
        self.assertIn("branch_priority_rows", ctx)
        self.assertIn("branch_supply_rows", ctx)
        self.assertEqual(ctx["plan_id"], self.plan.id)
        self.assertEqual(ctx["summary_label"], "Con bloqueos")
        self.assertEqual(ctx["stage_label"], "Recepciones bloqueadas ERP")
        self.assertIn("estatus=BLOCKED_ERP", ctx["next_action"]["url"])
        self.assertEqual(ctx["pipeline_steps"][2]["label"], "Recepciones")
        self.assertEqual(ctx["pipeline_steps"][2]["semaphore_label"], "Rojo")
        self.assertEqual(ctx["pipeline_steps"][2]["blocked"], 1)
        self.assertTrue(ctx["pipeline_steps"][2]["is_active"])
        self.assertEqual(ctx["pipeline_steps"][2]["action_label"], "Resolver recepciones")
        self.assertTrue(ctx["master_focus_rows"])
        focus_row = ctx["master_focus_rows"][0]
        self.assertIn(f"insumo_id={self.insumo.id}", focus_row["action_url"])
        self.assertEqual(focus_row["edit_url"], reverse("maestros:insumo_update", args=[self.insumo.id]))
        self.assertContains(response, "Editar artículo")
        self.assertContains(response, "Dependencias upstream del plan")
        self.assertIn("upstream_dependency_rows", ctx)
        self.assertEqual([row["label"] for row in ctx["upstream_dependency_rows"]], ["Plan de producción", "Maestro de artículos", "Órdenes sin bloqueo"])
        self.assertIn("Atiende bloqueos ERP o diferencias", ctx["pipeline_steps"][2]["action_detail"])
        self.assertEqual(ctx["stage_focus"]["label"], "Recepciones")
        self.assertEqual(len(ctx["stage_focus"]["rows"]), 1)
        self.assertEqual(ctx["closure_checks"][2]["label"], "Recepciones aplicadas")
        self.assertEqual(ctx["closure_checks"][2]["action_label"], "Resolver recepciones bloqueadas")
        self.assertIn("closure_key=recepciones_aplicadas", ctx["closure_checks"][2]["action_url"])
        self.assertIn("Aplica correcciones ERP", ctx["closure_checks"][2]["action_detail"])
        self.assertEqual(ctx["closure_focus"]["label"], "Recepciones aplicadas")
        self.assertGreaterEqual(len(ctx["closure_focus_rows"]), 1)
        self.assertEqual(ctx["closure_focus_rows"][0]["scope"], "Recepción")
        self.assertEqual(ctx["handoff_checks"][2]["label"], "Recepción → Cierre")
        self.assertEqual(ctx["handoff_checks"][2]["action_label"], "Resolver recepciones bloqueadas")
        self.assertIn("handoff_key=recepcion_cierre", ctx["handoff_checks"][2]["action_url"])
        self.assertEqual(ctx["handoff_focus"]["label"], "Recepción → Cierre")
        self.assertGreaterEqual(len(ctx["handoff_focus_rows"]), 1)
        self.assertEqual(ctx["handoff_focus_rows"][0]["scope"], "Recepción")
        self.assertIn("master_focus", ctx)
        self.assertEqual(ctx["master_focus"]["class_label"], "Materia prima")
        self.assertGreaterEqual(len(ctx["master_focus_rows"]), 1)
        self.assertContains(response, self.plan.nombre)
        self.assertContains(response, f"source=plan&amp;plan_id={self.plan.id}")
        self.assertContains(response, "bloqueados 1")
        self.assertContains(response, "Cierre ")
        self.assertContains(response, "erp-progress")
        self.assertContains(response, "Bloqueo prioritario por etapa")
        self.assertContains(response, "Resolver recepciones")
        self.assertEqual(list(response.context["ordenes"]), [self.orden_enviada])
        self.assertNotIn(other_orden, list(response.context["ordenes"]))
        recepcion = list(response.context["recepciones"])[0]
        self.assertEqual(recepcion.source_tipo, "plan")
        self.assertEqual(recepcion.source_plan_id, self.plan.id)
        self.assertEqual(recepcion.source_plan_nombre, self.plan.nombre)
        self.assertContains(response, "Bloqueos del plan")
        self.assertContains(response, self.plan.nombre)
        self.assertContains(response, "Recepciones bloqueadas")
        self.assertContains(response, "Bloqueo maestro prioritario")
        self.assertContains(response, "<th>Ámbito</th>", html=False)
        self.assertContains(response, "Resolver recepciones bloqueadas")
        self.assertContains(response, "Aplica correcciones ERP antes de cerrar recepción.")
        self.assertContains(response, "El cierre del plan sigue abierto por: recepciones aplicadas.")
        self.assertContains(response, "Etapa documental actual")
        self.assertContains(response, "Recepciones aplicadas")

    def test_orden_plan_scope_post_uses_plan_scope_reference(self):
        self.solicitud.area = f"PLAN_PRODUCCION:{self.plan.id}"
        self.solicitud.save(update_fields=["area"])

        response = self.client.post(
            reverse("compras:ordenes"),
            {
                "source": "plan",
                "plan_id": str(self.plan.id),
                "proveedor_id": str(self.proveedor_a.id),
                "solicitud_id": str(self.solicitud.id),
                "fecha_emision": date(2026, 2, 20).isoformat(),
                "estatus": OrdenCompra.STATUS_BORRADOR,
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("source=plan", response["Location"])
        self.assertIn(f"plan_id={self.plan.id}", response["Location"])

        orden = OrdenCompra.objects.order_by("-id").first()
        self.assertEqual(orden.solicitud_id, self.solicitud.id)
        self.assertEqual(orden.referencia, f"PLAN_PRODUCCION:{self.plan.id}")

    def test_orden_plan_scope_rejects_foreign_solicitud(self):
        initial_count = OrdenCompra.objects.count()
        response = self.client.post(
            reverse("compras:ordenes"),
            {
                "source": "plan",
                "plan_id": str(self.plan.id),
                "proveedor_id": str(self.proveedor_a.id),
                "solicitud_id": str(self.solicitud.id),
                "fecha_emision": date(2026, 2, 20).isoformat(),
                "estatus": OrdenCompra.STATUS_BORRADOR,
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(OrdenCompra.objects.count(), initial_count)
        self.assertContains(response, "no pertenece al plan")

    def test_recepcion_plan_scope_rejects_foreign_orden(self):
        foreign_order = OrdenCompra.objects.create(
            solicitud=self.solicitud,
            proveedor=self.proveedor_a,
            referencia="COMPRA GENERAL",
            fecha_emision=date(2026, 2, 22),
            monto_estimado=Decimal("80"),
            estatus=OrdenCompra.STATUS_CONFIRMADA,
        )
        initial_count = RecepcionCompra.objects.count()

        response = self.client.post(
            reverse("compras:recepciones"),
            {
                "source": "plan",
                "plan_id": str(self.plan.id),
                "orden_id": str(foreign_order.id),
                "fecha_recepcion": date(2026, 2, 23).isoformat(),
                "conformidad_pct": "100",
                "estatus": RecepcionCompra.STATUS_PENDIENTE,
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(RecepcionCompra.objects.count(), initial_count)
        self.assertContains(response, "no pertenece al plan")

    def test_ordenes_view_marks_reabasto_cedis_source(self):
        self.plan.nombre = "CEDIS Reabasto Matriz"
        self.plan.notas = "[AUTO_REABASTO_CEDIS:MATRIZ]"
        self.plan.save(update_fields=["nombre", "notas"])
        self.orden_enviada.referencia = f"PLAN_PRODUCCION:{self.plan.id}"
        self.orden_enviada.save(update_fields=["referencia"])
        self.solicitud.area = f"PLAN_PRODUCCION:{self.plan.id}"
        self.solicitud.save(update_fields=["area"])
        response = self.client.get(reverse("compras:ordenes"), {"q": f"PLAN_PRODUCCION:{self.plan.id}"})
        self.assertEqual(response.status_code, 200)
        orden = list(response.context["ordenes"])[0]
        self.assertEqual(orden.source_tipo, "reabasto_cedis")
        self.assertContains(response, "Reabasto CEDIS")

    def test_ordenes_view_filters_by_source_plan(self):
        self.orden_enviada.referencia = f"PLAN_PRODUCCION:{self.plan.id}"
        self.orden_enviada.save(update_fields=["referencia"])
        self.solicitud.area = f"PLAN_PRODUCCION:{self.plan.id}"
        self.solicitud.save(update_fields=["area"])
        response = self.client.get(reverse("compras:ordenes"), {"source": "plan", "plan_id": str(self.plan.id)})
        self.assertEqual(response.status_code, 200)
        ordenes = list(response.context["ordenes"])
        self.assertIn(self.orden_enviada, ordenes)
        self.assertIn(self.orden_cerrada, ordenes)
        orden_plan = next(o for o in ordenes if o.id == self.orden_enviada.id)
        self.assertEqual(orden_plan.source_tipo, "plan")
        self.assertContains(response, "Plan producción")

    def test_recepciones_view_marks_reabasto_cedis_source(self):
        self.plan.nombre = "CEDIS Reabasto Matriz"
        self.plan.notas = "[AUTO_REABASTO_CEDIS:MATRIZ]"
        self.plan.save(update_fields=["nombre", "notas"])
        self.orden_enviada.referencia = f"PLAN_PRODUCCION:{self.plan.id}"
        self.orden_enviada.save(update_fields=["referencia"])
        response = self.client.get(reverse("compras:recepciones"), {"q": f"PLAN_PRODUCCION:{self.plan.id}"})
        self.assertEqual(response.status_code, 200)
        recepcion = list(response.context["recepciones"])[0]
        self.assertEqual(recepcion.source_tipo, "reabasto_cedis")
        self.assertContains(response, "Reabasto CEDIS")

    def test_recepciones_view_filters_by_source_plan(self):
        self.orden_enviada.referencia = f"PLAN_PRODUCCION:{self.plan.id}"
        self.orden_enviada.save(update_fields=["referencia"])
        response = self.client.get(reverse("compras:recepciones"), {"source": "plan", "plan_id": str(self.plan.id)})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context["recepciones"]), [self.recepcion_pendiente])
        self.assertContains(response, "Plan producción")

    def test_recepciones_export_xlsx_respects_filters(self):
        self.orden_enviada.referencia = f"PLAN_PRODUCCION:{self.plan.id}"
        self.orden_enviada.save(update_fields=["referencia"])
        response = self.client.get(
            reverse("compras:recepciones"),
            {"estatus": RecepcionCompra.STATUS_PENDIENTE, "export": "xlsx"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        wb = load_workbook(BytesIO(response.content), data_only=True)
        ws = wb.active
        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        self.assertIn("origen", headers)
        self.assertIn("plan_origen", headers)
        values = [row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row and row[0]]
        self.assertIn(self.recepcion_pendiente.folio, values)
        self.assertNotIn(self.recepcion_cerrada.folio, values)
        data_rows = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        self.assertTrue(data_rows)
        self.assertEqual(data_rows[0][2], "plan")

    def test_recepcion_cerrada_aplica_entrada_a_inventario(self):
        existencia, _ = ExistenciaInsumo.objects.get_or_create(insumo=self.insumo)
        self.assertEqual(existencia.stock_actual, Decimal("0"))

        response = self.client.post(
            reverse("compras:recepciones"),
            {
                "orden_id": self.orden_enviada.id,
                "fecha_recepcion": "2026-02-22",
                "conformidad_pct": "100",
                "estatus": RecepcionCompra.STATUS_CERRADA,
                "observaciones": "cierre test",
            },
        )
        self.assertEqual(response.status_code, 302)

        recepcion = RecepcionCompra.objects.filter(orden=self.orden_enviada).order_by("-id").first()
        self.assertIsNotNone(recepcion)
        existencia.refresh_from_db()
        self.assertEqual(existencia.stock_actual, Decimal("2"))
        movimiento = MovimientoInventario.objects.get(source_hash=f"recepcion:{recepcion.id}:entrada")
        self.assertEqual(movimiento.tipo, MovimientoInventario.TIPO_ENTRADA)
        self.assertEqual(movimiento.cantidad, Decimal("2"))
        self.orden_enviada.refresh_from_db()
        self.assertEqual(self.orden_enviada.estatus, OrdenCompra.STATUS_CERRADA)

    def test_cerrar_recepcion_desde_estatus_aplica_entrada_una_sola_vez(self):
        existencia, _ = ExistenciaInsumo.objects.get_or_create(insumo=self.insumo)
        self.assertEqual(existencia.stock_actual, Decimal("0"))

        url = reverse(
            "compras:recepcion_estatus",
            kwargs={"pk": self.recepcion_pendiente.id, "estatus": RecepcionCompra.STATUS_CERRADA},
        )
        response = self.client.post(url)
        self.assertEqual(response.status_code, 302)
        self.recepcion_pendiente.refresh_from_db()
        self.assertEqual(self.recepcion_pendiente.estatus, RecepcionCompra.STATUS_CERRADA)
        existencia.refresh_from_db()
        self.assertEqual(existencia.stock_actual, Decimal("2"))
        self.assertEqual(
            MovimientoInventario.objects.filter(source_hash=f"recepcion:{self.recepcion_pendiente.id}:entrada").count(),
            1,
        )

    def test_recepcion_con_diferencias_sin_observaciones_muestra_bloqueo_erp(self):
        recepcion = RecepcionCompra.objects.create(
            orden=self.orden_enviada,
            fecha_recepcion=date(2026, 2, 23),
            conformidad_pct=Decimal("95"),
            estatus=RecepcionCompra.STATUS_DIFERENCIAS,
            observaciones="",
        )
        response = self.client.get(reverse("compras:recepciones"))
        self.assertEqual(response.status_code, 200)
        rendered = next(r for r in response.context["recepciones"] if r.id == recepcion.id)
        self.assertTrue(rendered.has_workflow_blockers)
        self.assertIn("Sin observaciones de diferencia", rendered.workflow_blockers)
        self.assertContains(response, "Bloqueada ERP")

    def test_no_cierra_recepcion_si_tiene_bloqueos_enterprise(self):
        recepcion = RecepcionCompra.objects.create(
            orden=self.orden_enviada,
            fecha_recepcion=date(2026, 2, 23),
            conformidad_pct=Decimal("95"),
            estatus=RecepcionCompra.STATUS_DIFERENCIAS,
            observaciones="",
        )
        response = self.client.post(
            reverse("compras:recepcion_estatus", kwargs={"pk": recepcion.id, "estatus": RecepcionCompra.STATUS_CERRADA}),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        recepcion.refresh_from_db()
        self.assertEqual(recepcion.estatus, RecepcionCompra.STATUS_DIFERENCIAS)
        self.assertContains(response, "No puedes mover")

    def test_recepciones_view_puede_filtrar_bloqueadas_erp(self):
        RecepcionCompra.objects.create(
            orden=self.orden_enviada,
            fecha_recepcion=date(2026, 2, 24),
            conformidad_pct=Decimal("95"),
            estatus=RecepcionCompra.STATUS_DIFERENCIAS,
            observaciones="",
        )
        response = self.client.get(reverse("compras:recepciones"), {"estatus": "BLOCKED_ERP"})
        self.assertEqual(response.status_code, 200)
        recepciones = list(response.context["recepciones"])
        self.assertTrue(recepciones)
        self.assertTrue(all(r.has_workflow_blockers for r in recepciones))

    def test_crear_orden_desde_solicitud_usa_proveedor_y_costo_canonico(self):
        canonical = Insumo.objects.create(
            nombre="Harina OC Canonica",
            categoria="Masa",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor_b,
            activo=True,
            codigo_point="OC-CAN-001",
        )
        variant = Insumo.objects.create(
            nombre="HARINA OC CANONICA",
            categoria="Masa",
            unidad_base=self.unidad,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=canonical,
            proveedor=self.proveedor_b,
            costo_unitario=Decimal("11.25"),
            source_hash="cost-oc-canonical",
        )
        solicitud = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=variant,
            cantidad=Decimal("3"),
            fecha_requerida=date(2026, 2, 24),
            estatus=SolicitudCompra.STATUS_APROBADA,
        )

        response = self.client.post(reverse("compras:solicitud_crear_orden", args=[solicitud.id]))
        self.assertEqual(response.status_code, 302)

        orden = OrdenCompra.objects.filter(solicitud=solicitud).order_by("-id").first()
        self.assertIsNotNone(orden)
        self.assertEqual(orden.proveedor_id, self.proveedor_b.id)
        self.assertEqual(orden.monto_estimado, Decimal("33.75"))

    def test_crear_orden_desde_solicitud_en_plan_preserva_scope_documental(self):
        insumo_plan = Insumo.objects.create(
            nombre="Harina plan documental",
            categoria="Masa",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor_a,
            activo=True,
            codigo_point="PLAN-DOC-001",
        )
        CostoInsumo.objects.create(
            insumo=insumo_plan,
            proveedor=self.proveedor_a,
            costo_unitario=Decimal("12.50"),
            source_hash="cost-plan-doc-001",
        )
        solicitud = SolicitudCompra.objects.create(
            area=f"PLAN_PRODUCCION:{self.plan.id}",
            solicitante="admin",
            insumo=insumo_plan,
            proveedor_sugerido=self.proveedor_a,
            cantidad=Decimal("3"),
            fecha_requerida=date(2026, 2, 24),
            estatus=SolicitudCompra.STATUS_APROBADA,
        )

        response = self.client.post(
            reverse("compras:solicitud_crear_orden", args=[solicitud.id]),
            {
                "source": "plan",
                "plan_id": str(self.plan.id),
                "return_query": f"source=plan&plan_id={self.plan.id}",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("source=plan", response["Location"])
        self.assertIn(f"plan_id={self.plan.id}", response["Location"])

        orden = OrdenCompra.objects.filter(solicitud=solicitud).order_by("-id").first()
        self.assertIsNotNone(orden)
        self.assertEqual(orden.referencia, f"PLAN_PRODUCCION:{self.plan.id}")

    def test_no_crea_oc_desde_plan_con_demanda_critica_bloqueada_por_maestro(self):
        insumo_plan = Insumo.objects.create(
            nombre="Harina plan bloqueada",
            categoria="",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor_a,
            activo=True,
        )
        receta_plan = Receta.objects.create(
            nombre="Base plan bloqueada",
            hash_contenido="test-hash-plan-blocked-001",
        )
        LineaReceta.objects.create(
            receta=receta_plan,
            posicion=1,
            insumo=insumo_plan,
            insumo_texto=insumo_plan.nombre,
            cantidad=Decimal("2"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("10"),
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100,
            match_status=LineaReceta.STATUS_AUTO,
        )
        PlanProduccionItem.objects.create(
            plan=self.plan,
            receta=receta_plan,
            cantidad=Decimal("1"),
        )
        sucursal = Sucursal.objects.create(codigo="SUC-PLAN-OC", nombre="Sucursal Plan OC", activa=True)
        VentaHistorica.objects.create(
            receta=receta_plan,
            sucursal=sucursal,
            fecha=self.plan.fecha_produccion - timedelta(days=2),
            cantidad=Decimal("95"),
            tickets=8,
        )
        solicitud = SolicitudCompra.objects.create(
            area=f"PLAN_PRODUCCION:{self.plan.id}",
            solicitante="admin",
            insumo=insumo_plan,
            proveedor_sugerido=self.proveedor_a,
            cantidad=Decimal("3"),
            fecha_requerida=date(2026, 2, 24),
            estatus=SolicitudCompra.STATUS_APROBADA,
        )

        response = self.client.post(
            reverse("compras:solicitud_crear_orden", args=[solicitud.id]),
            {
                "source": "plan",
                "plan_id": str(self.plan.id),
                "return_query": f"source=plan&plan_id={self.plan.id}",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(OrdenCompra.objects.filter(solicitud=solicitud).exists())
        self.assertContains(response, "No puedes emitir la orden de compra")

    def test_no_aprueba_solicitud_incompleta_por_bloqueo_enterprise(self):
        insumo_incompleto = Insumo.objects.create(
            nombre="Insumo sin proveedor ni codigo",
            categoria="Masa",
            unidad_base=self.unidad,
            activo=True,
        )
        solicitud = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=insumo_incompleto,
            cantidad=Decimal("2"),
            fecha_requerida=date(2026, 2, 24),
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )

        response = self.client.post(
            reverse("compras:solicitud_estatus", args=[solicitud.id, SolicitudCompra.STATUS_APROBADA]),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        solicitud.refresh_from_db()
        self.assertEqual(solicitud.estatus, SolicitudCompra.STATUS_BORRADOR)
        self.assertContains(response, "No puedes mover")

    def test_no_crea_oc_desde_solicitud_con_bloqueos_enterprise(self):
        canonical = Insumo.objects.create(
            nombre="Articulo sin costo vigente",
            categoria="Masa",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor_a,
            activo=True,
            codigo_point="NO-COST-001",
        )
        solicitud = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=canonical,
            cantidad=Decimal("3"),
            fecha_requerida=date(2026, 2, 24),
            estatus=SolicitudCompra.STATUS_APROBADA,
        )

        response = self.client.post(
            reverse("compras:solicitud_crear_orden", args=[solicitud.id]),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(OrdenCompra.objects.filter(solicitud=solicitud).exists())
        self.assertContains(response, "No puedes emitir la orden de compra")

    def test_recepcion_desde_variante_historica_aplica_inventario_al_canonico(self):
        canonical = Insumo.objects.create(
            nombre="Azucar Recepcion Canonica",
            categoria="Masa",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor_a,
            activo=True,
            codigo_point="REC-CAN-001",
        )
        variant = Insumo.objects.create(
            nombre="AZUCAR RECEPCION CANONICA",
            categoria="Masa",
            unidad_base=self.unidad,
            activo=True,
        )
        solicitud = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=variant,
            proveedor_sugerido=self.proveedor_a,
            cantidad=Decimal("4"),
            fecha_requerida=date(2026, 2, 25),
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        orden = OrdenCompra.objects.create(
            solicitud=solicitud,
            referencia="OC VARIANT RECEPCION",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 25),
            monto_estimado=Decimal("20"),
            estatus=OrdenCompra.STATUS_ENVIADA,
        )

        response = self.client.post(
            reverse("compras:recepciones"),
            {
                "orden_id": orden.id,
                "fecha_recepcion": "2026-02-26",
                "conformidad_pct": "100",
                "estatus": RecepcionCompra.STATUS_CERRADA,
                "observaciones": "cierre canonico",
            },
        )
        self.assertEqual(response.status_code, 302)

        canonical_existencia = ExistenciaInsumo.objects.get(insumo=canonical)
        self.assertEqual(canonical_existencia.stock_actual, Decimal("4"))
        self.assertFalse(ExistenciaInsumo.objects.filter(insumo=variant).exists())

        recepcion = RecepcionCompra.objects.filter(orden=orden).order_by("-id").first()
        self.assertIsNotNone(recepcion)
        movimiento = MovimientoInventario.objects.get(source_hash=f"recepcion:{recepcion.id}:entrada")
        self.assertEqual(movimiento.insumo_id, canonical.id)
        self.assertEqual(movimiento.cantidad, Decimal("4"))

    def test_ordenes_view_exposes_workflow_summary_and_stage(self):
        response = self.client.get(reverse("compras:ordenes"))
        self.assertEqual(response.status_code, 200)
        workflow_summary = response.context["workflow_summary"]
        cards = {item["label"]: item["count"] for item in workflow_summary["cards"]}
        self.assertEqual(cards["Enviadas"], 1)
        self.assertIn("gate_cards", workflow_summary)
        self.assertTrue(any(item["key"] == "ready_for_recepcion" for item in workflow_summary["gate_cards"]))
        self.assertIn("release_gate_rows", response.context)
        self.assertIn("release_gate_completion", response.context)
        self.assertContains(response, "Resumen de compras")
        self.assertContains(response, "Control Documental")
        self.assertContains(response, "Dependencias del flujo")
        self.assertContains(response, "Solicitudes / BOM")
        self.assertContains(response, "Órdenes documentales")
        self.assertContains(response, "Recepciones / Inventario")
        self.assertContains(response, "Criterios de cierre")
        self.assertContains(response, "Prioridades de atención")
        self.assertContains(response, "Resumen de seguimiento")
        self.assertContains(response, "Control por frente")
        self.assertContains(response, "Cierre global")
        self.assertIn("erp_command_center", response.context)
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertIn("trunk_handoff_rows", response.context)
        self.assertEqual(len(response.context["trunk_handoff_rows"]), 3)
        self.assertContains(response, "Semáforo de etapa")
        self.assertContains(response, "ERP completo")
        self.assertContains(response, "Esperando confirmación")
        self.assertContains(response, "Confirmar o marcar parcial")

    def test_ordenes_view_exposes_enterprise_board(self):
        self.solicitud.area = f"PLAN_PRODUCCION:{self.plan.id}"
        self.solicitud.save(update_fields=["area"])
        insumo_incompleto = Insumo.objects.create(
            nombre="Harina sin point orden",
            categoria="Masa",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor_a,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=insumo_incompleto,
            proveedor=self.proveedor_a,
            costo_unitario=Decimal("22"),
            source_hash="cost-harina-orden-master",
        )
        solicitud_bloqueada = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=insumo_incompleto,
            cantidad=Decimal("3"),
            fecha_requerida=date(2026, 2, 20),
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        orden_lista = OrdenCompra.objects.create(
            solicitud=self.solicitud,
            referencia=f"PLAN_PRODUCCION:{self.plan.id}",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 20),
            fecha_entrega_estimada=date(2026, 2, 22),
            monto_estimado=Decimal("150"),
            estatus=OrdenCompra.STATUS_CONFIRMADA,
        )
        orden_bloqueada = OrdenCompra.objects.create(
            solicitud=solicitud_bloqueada,
            referencia="OC BLOQUEADA TABLERO",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 20),
            fecha_entrega_estimada=None,
            monto_estimado=Decimal("0"),
            estatus=OrdenCompra.STATUS_BORRADOR,
        )

        response = self.client.get(reverse("compras:ordenes"))
        self.assertEqual(response.status_code, 200)
        board = response.context["enterprise_board"]
        next_steps = {item["label"]: item["count"] for item in board["next_step_cards"]}
        blockers = {item["label"]: item["count"] for item in board["blocker_cards"]}
        sources = {item["label"]: item["count"] for item in board["source_cards"]}
        self.assertGreaterEqual(next_steps["Registrar recepción"], 1)
        self.assertGreaterEqual(next_steps["Corrección ERP"], 1)
        self.assertGreaterEqual(blockers["Monto en cero"], 1)
        self.assertGreaterEqual(sources["Manual"], 1)
        self.assertGreaterEqual(sources["Plan producción"], 1)
        self.assertIn("master_blocker_class_cards", board)
        self.assertIn("master_blocker_missing_cards", board)
        self.assertIn("master_blocker_detail_rows", board)
        self.assertTrue(any(item["class_label"] == "Materia prima" for item in board["master_blocker_class_cards"]))
        self.assertTrue(any(item["key"] == "codigo_point" for item in board["master_blocker_missing_cards"]))
        self.assertTrue(any(item["class_label"] == "Materia prima" for item in board["master_blocker_detail_rows"]))
        self.assertTrue(
            any(
                item["key"] == "registrar_recepcion" and item["query"] == "?workflow_action=registrar_recepcion"
                for item in board["next_step_cards"]
            )
        )
        self.assertTrue(
            any(
                item["key"] == "monto_cero" and item["query"] == "?blocker_key=monto_cero"
                for item in board["blocker_cards"]
            )
        )
        self.assertContains(response, "Salud operativa ERP")
        self.assertContains(response, "Plan producción")
        self.assertContains(response, "Bloqueada ERP")
        self.assertContains(response, "Bloqueos del maestro por clase")
        self.assertContains(response, "Bloqueos del maestro por dato faltante")
        self.assertContains(response, "Detalle del maestro bloqueando órdenes")
        self.assertContains(response, "Maestro ERP")
        self.assertContains(response, "Incompleto")
        self.assertContains(response, "Harina sin point orden")
        self.assertContains(response, "código Point")
        self.assertContains(response, "Registrar código comercial")
        self.assertContains(response, "Captura el código comercial")
        self.assertContains(response, orden_lista.folio)
        self.assertContains(response, f"insumo_id={insumo_incompleto.id}")
        self.assertContains(response, reverse("maestros:insumo_update", args=[insumo_incompleto.id]))
        self.assertContains(response, "Editar artículo")
        self.assertContains(response, orden_bloqueada.folio)
        self.assertContains(response, 'Enviar</button>')
        self.assertContains(response, 'disabled title="')

    def test_ordenes_view_can_filter_by_workflow_action_and_blocker_key(self):
        orden_bloqueada = OrdenCompra.objects.create(
            solicitud=self.solicitud,
            referencia="OC WF BLOQ",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 20),
            fecha_entrega_estimada=None,
            monto_estimado=Decimal("0"),
            estatus=OrdenCompra.STATUS_BORRADOR,
        )
        orden_lista = OrdenCompra.objects.create(
            solicitud=self.solicitud,
            referencia="OC WF LISTA",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 20),
            fecha_entrega_estimada=date(2026, 2, 22),
            monto_estimado=Decimal("120"),
            estatus=OrdenCompra.STATUS_CONFIRMADA,
        )

        response = self.client.get(reverse("compras:ordenes"), {"workflow_action": "registrar_recepcion"})
        self.assertContains(response, orden_lista.folio)
        self.assertNotContains(response, orden_bloqueada.folio)

        response = self.client.get(reverse("compras:ordenes"), {"blocker_key": "monto_cero"})
        self.assertContains(response, orden_bloqueada.folio)
        self.assertNotContains(response, orden_lista.folio)

    def test_ordenes_view_can_filter_by_closure_and_handoff_key(self):
        orden_bloqueada = OrdenCompra.objects.create(
            solicitud=self.solicitud,
            referencia=f"PLAN_PRODUCCION:{self.plan.id}",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 20),
            fecha_entrega_estimada=None,
            monto_estimado=Decimal("0"),
            estatus=OrdenCompra.STATUS_BORRADOR,
        )
        orden_confirmada = OrdenCompra.objects.create(
            solicitud=self.solicitud,
            referencia=f"PLAN_PRODUCCION:{self.plan.id}",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 20),
            fecha_entrega_estimada=date(2026, 2, 22),
            monto_estimado=Decimal("120"),
            estatus=OrdenCompra.STATUS_CONFIRMADA,
        )

        response = self.client.get(
            reverse("compras:ordenes"),
            {"source": "plan", "plan_id": str(self.plan.id), "closure_key": "ordenes_sin_bloqueo"},
        )
        self.assertContains(response, orden_bloqueada.folio)
        self.assertNotContains(response, orden_confirmada.folio)
        self.assertEqual(response.context["closure_key_filter"], "ordenes_sin_bloqueo")

        response = self.client.get(
            reverse("compras:ordenes"),
            {"source": "plan", "plan_id": str(self.plan.id), "handoff_key": "orden_recepcion"},
        )
        self.assertContains(response, orden_bloqueada.folio)
        self.assertContains(response, orden_confirmada.folio)
        self.assertEqual(response.context["handoff_key_filter"], "orden_recepcion")

    def test_ordenes_view_shows_blocker_summary(self):
        orden = OrdenCompra.objects.create(
            solicitud=self.solicitud,
            referencia="OC BLOQUEO COMPACTO",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 20),
            fecha_entrega_estimada=None,
            monto_estimado=Decimal("0"),
            estatus=OrdenCompra.STATUS_BORRADOR,
        )

        response = self.client.get(reverse("compras:ordenes"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Monto estimado en cero")
        self.assertContains(response, "Sin fecha de entrega estimada")
        self.assertContains(response, "Corregir datos ERP")
        self.assertContains(response, orden.folio)
        self.assertContains(response, "Abrir orden")
        self.assertContains(response, "Corregir monto")
        self.assertContains(response, "Registrar entrega")
        self.assertContains(response, "Orden bloqueada")
        self.assertContains(response, f"estatus=BLOCKED_ERP&amp;q={orden.folio}")

    def test_ordenes_view_can_filter_by_master_blocker(self):
        insumo_mp = Insumo.objects.create(
            nombre="Harina sin point filtro orden",
            categoria="Masa",
            unidad_base=self.unidad,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=insumo_mp,
            proveedor=self.proveedor_a,
            costo_unitario=Decimal("20"),
            source_hash="cost-mp-filtro-orden",
        )
        solicitud_mp = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=insumo_mp,
            cantidad=Decimal("2"),
            fecha_requerida=date(2026, 2, 20),
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        orden_mp = OrdenCompra.objects.create(
            solicitud=solicitud_mp,
            referencia="OC FILTRO MP",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 20),
            monto_estimado=Decimal("10"),
            estatus=OrdenCompra.STATUS_BORRADOR,
        )
        insumo_pack = Insumo.objects.create(
            nombre="Etiqueta sin point filtro orden",
            categoria="Empaques",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor_a,
            activo=True,
            tipo_item=Insumo.TIPO_EMPAQUE,
        )
        CostoInsumo.objects.create(
            insumo=insumo_pack,
            proveedor=self.proveedor_a,
            costo_unitario=Decimal("2"),
            source_hash="cost-pack-filtro-orden",
        )
        solicitud_pack = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=insumo_pack,
            cantidad=Decimal("4"),
            fecha_requerida=date(2026, 2, 20),
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        OrdenCompra.objects.create(
            solicitud=solicitud_pack,
            referencia="OC FILTRO EMPAQUE",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 20),
            monto_estimado=Decimal("10"),
            estatus=OrdenCompra.STATUS_BORRADOR,
        )

        response = self.client.get(
            reverse("compras:ordenes"),
            {"master_class": Insumo.TIPO_MATERIA_PRIMA, "master_missing": "proveedor"},
        )
        self.assertEqual(response.status_code, 200)
        rendered_folios = [item.folio for item in response.context["ordenes"]]
        self.assertIn(orden_mp.folio, rendered_folios)
        self.assertNotIn("OC FILTRO EMPAQUE", rendered_folios)
        self.assertEqual(response.context["master_class_filter"], Insumo.TIPO_MATERIA_PRIMA)
        self.assertEqual(response.context["master_missing_filter"], "proveedor")

    def test_orden_borrador_con_datos_incompletos_muestra_bloqueo_erp(self):
        orden = OrdenCompra.objects.create(
            solicitud=self.solicitud,
            referencia="OC BLOQUEADA ERP",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 20),
            fecha_entrega_estimada=None,
            monto_estimado=Decimal("0"),
            estatus=OrdenCompra.STATUS_BORRADOR,
        )
        response = self.client.get(reverse("compras:ordenes"))
        self.assertEqual(response.status_code, 200)
        rendered = next(o for o in response.context["ordenes"] if o.id == orden.id)
        self.assertTrue(rendered.has_workflow_blockers)
        self.assertIn("Monto estimado en cero", rendered.workflow_blockers)
        self.assertContains(response, "Bloqueada ERP")

    def test_no_envia_orden_si_tiene_bloqueos_enterprise(self):
        orden = OrdenCompra.objects.create(
            solicitud=self.solicitud,
            referencia="OC NO ENVIABLE",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 20),
            fecha_entrega_estimada=None,
            monto_estimado=Decimal("0"),
            estatus=OrdenCompra.STATUS_BORRADOR,
        )
        response = self.client.post(
            reverse("compras:orden_estatus", args=[orden.id, OrdenCompra.STATUS_ENVIADA]),
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        orden.refresh_from_db()
        self.assertEqual(orden.estatus, OrdenCompra.STATUS_BORRADOR)
        self.assertContains(response, "No puedes mover")

    def test_recepciones_view_exposes_workflow_summary_and_stage(self):
        response = self.client.get(reverse("compras:recepciones"))
        self.assertEqual(response.status_code, 200)
        workflow_summary = response.context["workflow_summary"]
        cards = {item["label"]: item["count"] for item in workflow_summary["cards"]}
        self.assertEqual(cards["Por validar"], 1)
        self.assertIn("gate_cards", workflow_summary)
        self.assertTrue(any(item["key"] == "applied_inventory" for item in workflow_summary["gate_cards"]))
        self.assertIn("release_gate_rows", response.context)
        self.assertIn("release_gate_completion", response.context)
        self.assertContains(response, "Resumen de compras")
        self.assertContains(response, "Control Documental")
        self.assertContains(response, "Dependencias del flujo")
        self.assertContains(response, "Solicitudes / BOM")
        self.assertContains(response, "Órdenes documentales")
        self.assertContains(response, "Recepciones / Inventario")
        self.assertContains(response, "Criterios de cierre")
        self.assertContains(response, "Prioridades de atención")
        self.assertContains(response, "Resumen de seguimiento")
        self.assertContains(response, "Control por frente")
        self.assertContains(response, "Cierre global")
        self.assertIn("erp_command_center", response.context)
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertIn("trunk_handoff_rows", response.context)
        self.assertEqual(len(response.context["trunk_handoff_rows"]), 3)
        self.assertContains(response, "Semáforo de etapa")
        self.assertContains(response, "Recepción válida")
        self.assertContains(response, "Validación")
        self.assertContains(response, "Cerrar o marcar diferencias")

    def test_recepciones_view_exposes_enterprise_board(self):
        self.orden_enviada.referencia = f"PLAN_PRODUCCION:{self.plan.id}"
        self.orden_enviada.save(update_fields=["referencia"])
        insumo_incompleto = Insumo.objects.create(
            nombre="Etiqueta sin codigo recepcion",
            categoria="Empaques",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor_a,
            activo=True,
            tipo_item=Insumo.TIPO_EMPAQUE,
        )
        CostoInsumo.objects.create(
            insumo=insumo_incompleto,
            proveedor=self.proveedor_a,
            costo_unitario=Decimal("3"),
            source_hash="cost-etiqueta-recepcion-master",
        )
        solicitud_bloqueada = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=insumo_incompleto,
            cantidad=Decimal("10"),
            fecha_requerida=date(2026, 2, 20),
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        orden_bloqueada = OrdenCompra.objects.create(
            solicitud=solicitud_bloqueada,
            referencia="PLAN_PRODUCCION:MASTER-RECEPCION",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 20),
            fecha_entrega_estimada=date(2026, 2, 22),
            monto_estimado=Decimal("40"),
            estatus=OrdenCompra.STATUS_CONFIRMADA,
        )
        RecepcionCompra.objects.create(
            orden=self.orden_enviada,
            fecha_recepcion=date(2026, 2, 21),
            conformidad_pct=Decimal("100"),
            estatus=RecepcionCompra.STATUS_CERRADA,
            observaciones="ok",
        )
        recepcion_bloqueada = RecepcionCompra.objects.create(
            orden=orden_bloqueada,
            fecha_recepcion=date(2026, 2, 21),
            conformidad_pct=Decimal("50"),
            estatus=RecepcionCompra.STATUS_DIFERENCIAS,
            observaciones="",
        )

        response = self.client.get(reverse("compras:recepciones"))
        self.assertEqual(response.status_code, 200)
        board = response.context["enterprise_board"]
        next_steps = {item["label"]: item["count"] for item in board["next_step_cards"]}
        blockers = {item["label"]: item["count"] for item in board["blocker_cards"]}
        sources = {item["label"]: item["count"] for item in board["source_cards"]}
        self.assertGreaterEqual(next_steps["Corrección ERP"], 1)
        self.assertGreaterEqual(blockers["Sin observación"], 1)
        self.assertGreaterEqual(board["applied_total"], 1)
        self.assertGreaterEqual(sources["Plan producción"], 1)
        self.assertIn("master_blocker_class_cards", board)
        self.assertIn("master_blocker_missing_cards", board)
        self.assertIn("master_blocker_detail_rows", board)
        self.assertTrue(any(item["class_label"] == "Empaque" for item in board["master_blocker_class_cards"]))
        self.assertTrue(any(item["key"] == "codigo_point" for item in board["master_blocker_missing_cards"]))
        self.assertTrue(any(item["class_label"] == "Empaque" for item in board["master_blocker_detail_rows"]))
        self.assertTrue(
            any(
                item["key"] == "corregir_recepcion" and item["query"] == "?workflow_action=corregir_recepcion"
                for item in board["next_step_cards"]
            )
        )
        self.assertTrue(
            any(
                item["key"] == "sin_observacion" and item["query"] == "?blocker_key=sin_observacion"
                for item in board["blocker_cards"]
            )
        )
        self.assertContains(response, "Salud operativa ERP")
        self.assertContains(response, "Plan producción")
        self.assertContains(response, "Corregir datos recepción")
        self.assertContains(response, "Bloqueos del maestro por clase")
        self.assertContains(response, "Bloqueos del maestro por dato faltante")
        self.assertContains(response, "Detalle del maestro bloqueando recepciones")
        self.assertContains(response, "Maestro ERP")
        self.assertContains(response, "Incompleto")
        self.assertContains(response, "Etiqueta sin codigo recepcion")
        self.assertContains(response, "código Point")
        self.assertContains(response, "Registrar código comercial")
        self.assertContains(response, "Captura el código comercial")
        self.assertContains(response, recepcion_bloqueada.folio)
        self.assertContains(response, f"insumo_id={insumo_incompleto.id}")
        self.assertContains(response, reverse("maestros:insumo_update", args=[insumo_incompleto.id]))
        self.assertContains(response, "Editar artículo")

    def test_recepciones_view_can_filter_by_workflow_action_and_blocker_key(self):
        recepcion_lista = RecepcionCompra.objects.create(
            orden=self.orden_enviada,
            fecha_recepcion=date(2026, 2, 21),
            conformidad_pct=Decimal("100"),
            estatus=RecepcionCompra.STATUS_PENDIENTE,
            observaciones="ok",
        )
        recepcion_bloqueada = RecepcionCompra.objects.create(
            orden=self.orden_enviada,
            fecha_recepcion=date(2026, 2, 21),
            conformidad_pct=Decimal("50"),
            estatus=RecepcionCompra.STATUS_DIFERENCIAS,
            observaciones="",
        )

        response = self.client.get(reverse("compras:recepciones"), {"workflow_action": "cerrar_diferencias"})
        self.assertContains(response, recepcion_lista.folio)
        self.assertNotContains(response, recepcion_bloqueada.folio)

        response = self.client.get(reverse("compras:recepciones"), {"blocker_key": "sin_observacion"})
        self.assertContains(response, recepcion_bloqueada.folio)
        self.assertNotContains(response, recepcion_lista.folio)

    def test_recepciones_view_can_filter_by_closure_and_handoff_key(self):
        self.orden_enviada.referencia = f"PLAN_PRODUCCION:{self.plan.id}"
        self.orden_enviada.save(update_fields=["referencia"])
        recepcion_pendiente = RecepcionCompra.objects.create(
            orden=self.orden_enviada,
            fecha_recepcion=date(2026, 2, 21),
            conformidad_pct=Decimal("100"),
            estatus=RecepcionCompra.STATUS_PENDIENTE,
            observaciones="ok",
        )
        recepcion_bloqueada = RecepcionCompra.objects.create(
            orden=self.orden_enviada,
            fecha_recepcion=date(2026, 2, 21),
            conformidad_pct=Decimal("50"),
            estatus=RecepcionCompra.STATUS_DIFERENCIAS,
            observaciones="",
        )
        RecepcionCompra.objects.create(
            orden=self.orden_enviada,
            fecha_recepcion=date(2026, 2, 22),
            conformidad_pct=Decimal("100"),
            estatus=RecepcionCompra.STATUS_CERRADA,
            observaciones="ok",
        )

        response = self.client.get(
            reverse("compras:recepciones"),
            {"source": "plan", "plan_id": str(self.plan.id), "closure_key": "recepciones_aplicadas"},
        )
        self.assertContains(response, recepcion_pendiente.folio)
        self.assertContains(response, recepcion_bloqueada.folio)
        self.assertEqual(response.context["closure_key_filter"], "recepciones_aplicadas")

        response = self.client.get(
            reverse("compras:recepciones"),
            {"source": "plan", "plan_id": str(self.plan.id), "handoff_key": "recepcion_cierre"},
        )
        self.assertContains(response, recepcion_pendiente.folio)
        self.assertContains(response, recepcion_bloqueada.folio)
        self.assertEqual(response.context["handoff_key_filter"], "recepcion_cierre")

    def test_recepciones_view_shows_blocker_summary(self):
        recepcion = RecepcionCompra.objects.create(
            orden=self.orden_enviada,
            fecha_recepcion=date(2026, 2, 21),
            conformidad_pct=Decimal("50"),
            estatus=RecepcionCompra.STATUS_DIFERENCIAS,
            observaciones="",
        )

        response = self.client.get(reverse("compras:recepciones"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sin observaciones de diferencia")
        self.assertContains(response, "Corregir datos recepción")
        self.assertContains(response, "Bloqueada ERP")
        self.assertContains(response, recepcion.folio)
        self.assertContains(response, "Abrir recepción")
        self.assertContains(response, "Registrar observaciones")
        self.assertContains(response, "Recepción bloqueada")
        self.assertContains(response, f"estatus=BLOCKED_ERP&amp;q={recepcion.folio}")
        self.assertContains(response, 'Cerrar</button>')
        self.assertContains(response, 'disabled title="Sin observaciones de diferencia"')

    def test_recepciones_view_can_filter_by_master_blocker(self):
        insumo_mp = Insumo.objects.create(
            nombre="Harina sin point filtro recepcion",
            categoria="Masa",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor_a,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=insumo_mp,
            proveedor=self.proveedor_a,
            costo_unitario=Decimal("20"),
            source_hash="cost-mp-filtro-recepcion",
        )
        solicitud_mp = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=insumo_mp,
            cantidad=Decimal("2"),
            fecha_requerida=date(2026, 2, 20),
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        orden_mp = OrdenCompra.objects.create(
            solicitud=solicitud_mp,
            referencia="OC FILTRO REC MP",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 20),
            monto_estimado=Decimal("10"),
            estatus=OrdenCompra.STATUS_CONFIRMADA,
        )
        recepcion_mp = RecepcionCompra.objects.create(
            orden=orden_mp,
            fecha_recepcion=date(2026, 2, 21),
            conformidad_pct=Decimal("100"),
            estatus=RecepcionCompra.STATUS_PENDIENTE,
            observaciones="ok",
        )
        insumo_pack = Insumo.objects.create(
            nombre="Etiqueta sin point filtro recepcion",
            categoria="",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor_a,
            activo=True,
            tipo_item=Insumo.TIPO_EMPAQUE,
        )
        CostoInsumo.objects.create(
            insumo=insumo_pack,
            proveedor=self.proveedor_a,
            costo_unitario=Decimal("2"),
            source_hash="cost-pack-filtro-recepcion",
        )
        solicitud_pack = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=insumo_pack,
            cantidad=Decimal("4"),
            fecha_requerida=date(2026, 2, 20),
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        orden_pack = OrdenCompra.objects.create(
            solicitud=solicitud_pack,
            referencia="OC FILTRO REC EMPAQUE",
            proveedor=self.proveedor_a,
            fecha_emision=date(2026, 2, 20),
            monto_estimado=Decimal("10"),
            estatus=OrdenCompra.STATUS_CONFIRMADA,
        )
        recepcion_pack = RecepcionCompra.objects.create(
            orden=orden_pack,
            fecha_recepcion=date(2026, 2, 21),
            conformidad_pct=Decimal("100"),
            estatus=RecepcionCompra.STATUS_PENDIENTE,
            observaciones="ok",
        )

        response = self.client.get(
            reverse("compras:recepciones"),
            {"master_class": Insumo.TIPO_EMPAQUE, "master_missing": "categoria"},
        )
        self.assertEqual(response.status_code, 200)
        rendered_folios = [item.folio for item in response.context["recepciones"]]
        self.assertIn(recepcion_pack.folio, rendered_folios)
        self.assertNotIn(recepcion_mp.folio, rendered_folios)
        self.assertEqual(response.context["master_class_filter"], Insumo.TIPO_EMPAQUE)
        self.assertEqual(response.context["master_missing_filter"], "categoria")


class ComprasFolioRetryTests(TestCase):
    def setUp(self):
        self.proveedor = Proveedor.objects.create(nombre="Proveedor Folio", activo=True)
        self.unidad = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.insumo = Insumo.objects.create(
            nombre="Insumo Folio",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor,
            activo=True,
        )

    def test_solicitud_folio_retries_on_collision(self):
        SolicitudCompra.objects.create(
            folio="SOL-COLLIDE-001",
            area="Compras",
            solicitante="tester",
            insumo=self.insumo,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("1"),
            fecha_requerida=date(2026, 2, 26),
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        with patch.object(
            SolicitudCompra,
            "_next_folio",
            side_effect=["SOL-COLLIDE-001", "SOL-COLLIDE-002"],
        ):
            solicitud = SolicitudCompra.objects.create(
                area="Compras",
                solicitante="tester",
                insumo=self.insumo,
                proveedor_sugerido=self.proveedor,
                cantidad=Decimal("2"),
                fecha_requerida=date(2026, 2, 27),
                estatus=SolicitudCompra.STATUS_BORRADOR,
            )
        self.assertEqual(solicitud.folio, "SOL-COLLIDE-002")

    def test_orden_folio_retries_on_collision(self):
        solicitud = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="tester",
            insumo=self.insumo,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("1"),
            fecha_requerida=date(2026, 2, 26),
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        OrdenCompra.objects.create(
            folio="OC-COLLIDE-001",
            solicitud=solicitud,
            proveedor=self.proveedor,
            fecha_emision=date(2026, 2, 26),
            monto_estimado=Decimal("10"),
            estatus=OrdenCompra.STATUS_BORRADOR,
        )
        with patch.object(
            OrdenCompra,
            "_next_folio",
            side_effect=["OC-COLLIDE-001", "OC-COLLIDE-002"],
        ):
            orden = OrdenCompra.objects.create(
                solicitud=solicitud,
                proveedor=self.proveedor,
                fecha_emision=date(2026, 2, 27),
                monto_estimado=Decimal("12"),
                estatus=OrdenCompra.STATUS_BORRADOR,
            )
        self.assertEqual(orden.folio, "OC-COLLIDE-002")

    def test_recepcion_folio_retries_on_collision(self):
        solicitud = SolicitudCompra.objects.create(
            area="Compras",
            solicitante="tester",
            insumo=self.insumo,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("1"),
            fecha_requerida=date(2026, 2, 26),
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        orden = OrdenCompra.objects.create(
            solicitud=solicitud,
            proveedor=self.proveedor,
            fecha_emision=date(2026, 2, 26),
            monto_estimado=Decimal("10"),
            estatus=OrdenCompra.STATUS_BORRADOR,
        )
        RecepcionCompra.objects.create(
            folio="REC-COLLIDE-001",
            orden=orden,
            fecha_recepcion=date(2026, 2, 26),
            conformidad_pct=Decimal("100"),
            estatus=RecepcionCompra.STATUS_PENDIENTE,
            observaciones="",
        )
        with patch.object(
            RecepcionCompra,
            "_next_folio",
            side_effect=["REC-COLLIDE-001", "REC-COLLIDE-002"],
        ):
            recepcion = RecepcionCompra.objects.create(
                orden=orden,
                fecha_recepcion=date(2026, 2, 27),
                conformidad_pct=Decimal("100"),
                estatus=RecepcionCompra.STATUS_PENDIENTE,
                observaciones="",
            )
        self.assertEqual(recepcion.folio, "REC-COLLIDE-002")
