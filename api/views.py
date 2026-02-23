import csv
from collections import defaultdict
from contextlib import nullcontext
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from django.db import transaction, OperationalError, ProgrammingError
from django.db.models import Count, Max, Q, Sum
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from openpyxl import Workbook
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.authtoken.models import Token
from rest_framework.authtoken.views import ObtainAuthToken

from activos.models import Activo, BitacoraMantenimiento, OrdenMantenimiento, PlanMantenimiento
from control.models import MermaPOS, VentaPOS
from control.services import build_discrepancias_report, resolve_period_range
from compras.models import OrdenCompra, RecepcionCompra, SolicitudCompra
from core.access import (
    ROLE_ADMIN,
    ROLE_COMPRAS,
    ROLE_DG,
    can_manage_compras,
    can_manage_inventario,
    can_view_audit,
    can_view_compras,
    can_view_inventario,
    can_view_maestros,
    can_view_reportes,
    has_any_role,
)
from core.audit import log_event
from core.models import AuditLog, Sucursal
from compras.views import (
    _apply_recepcion_to_inventario,
    _active_solicitud_statuses,
    _build_budget_context,
    _build_budget_history,
    _build_category_dashboard,
    _can_transition_orden,
    _can_transition_recepcion,
    _can_transition_solicitud,
    _build_consumo_vs_plan_dashboard,
    _default_fecha_requerida,
    _parse_date_value,
    _build_provider_dashboard,
    _filtered_solicitudes,
    _resolve_proveedor_name,
    _sanitize_consumo_ref_filter,
)
from integraciones.models import PublicApiAccessLog, PublicApiClient
from integraciones.views import _deactivate_idle_api_clients, _purge_api_logs
from inventario.models import AjusteInventario, AlmacenSyncRun, ExistenciaInsumo
from inventario.views import (
    _apply_ajuste,
    _apply_cross_filters,
    _build_cross_unified_rows,
    _build_pending_grouped,
    _resolve_cross_source_with_alias,
)
from maestros.models import CostoInsumo, Insumo, InsumoAlias, PointPendingMatch, Proveedor
from recetas.models import (
    LineaReceta,
    PlanProduccion,
    PlanProduccionItem,
    PronosticoVenta,
    Receta,
    RecetaCodigoPointAlias,
    SolicitudVenta,
    VentaHistorica,
    normalizar_codigo_point,
)
from recetas.views import (
    _build_forecast_from_history,
    _filter_forecast_result_by_confianza,
    _forecast_session_payload,
    _forecast_vs_solicitud_preview,
    _normalize_periodo_mes,
    _resolve_receta_for_sales,
    _resolve_solicitud_window,
    _resolve_sucursal_for_sales,
    _ui_to_model_alcance,
)
from recetas.utils.normalizacion import normalizar_nombre
from recetas.utils.matching import match_insumo
from recetas.utils.costeo_versionado import asegurar_version_costeo, comparativo_versiones
from .serializers import (
    ComprasSolicitudImportConfirmSerializer,
    ComprasSolicitudImportPreviewSerializer,
    ComprasCrearOrdenSerializer,
    ComprasOrdenStatusSerializer,
    ComprasRecepcionCreateSerializer,
    ComprasRecepcionStatusSerializer,
    ComprasSolicitudCreateSerializer,
    ComprasSolicitudStatusSerializer,
    ControlMermaPosBulkSerializer,
    ControlVentaPosBulkSerializer,
    ActivosOrdenCreateSerializer,
    ActivosOrdenStatusSerializer,
    ForecastBacktestRequestSerializer,
    ForecastEstadisticoGuardarSerializer,
    ForecastEstadisticoRequestSerializer,
    IntegracionesDeactivateIdleClientsSerializer,
    IntegracionesMaintenanceRunSerializer,
    IntegracionesPurgeApiLogsSerializer,
    InventarioAjusteCreateSerializer,
    InventarioAjusteDecisionSerializer,
    InventarioAliasCreateSerializer,
    InventarioCrossPendientesResolveSerializer,
    InventarioAliasMassReassignSerializer,
    InventarioPointPendingResolveSerializer,
    MRPRequestSerializer,
    MRPRequerimientosRequestSerializer,
    PlanProduccionCreateSerializer,
    PlanProduccionItemCreateSerializer,
    PlanProduccionItemUpdateSerializer,
    PlanProduccionUpdateSerializer,
    PlanDesdePronosticoRequestSerializer,
    PronosticoVentaBulkSerializer,
    RecetaCostoVersionSerializer,
    SolicitudVentaAplicarForecastSerializer,
    SolicitudVentaBulkSerializer,
    SolicitudVentaUpsertSerializer,
    VentaHistoricaBulkSerializer,
)


def _load_versiones_costeo(receta: Receta, limit: int):
    return list(receta.versiones_costo.order_by("-version_num")[:limit])


def _to_decimal(value, default: Decimal = Decimal("0")) -> Decimal:
    if value is None:
        return default
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return default


def _parse_bool(value: str | None, default: bool = False) -> bool:
    if value is None:
        return default
    raw = str(value).strip().lower()
    if raw in {"1", "true", "yes", "si", "on"}:
        return True
    if raw in {"0", "false", "no", "off"}:
        return False
    return default


def _parse_period(period_raw: str | None) -> tuple[int, int] | None:
    if not period_raw:
        return None
    raw = str(period_raw).strip()
    parts = raw.split("-")
    if len(parts) != 2:
        return None
    try:
        year = int(parts[0])
        month = int(parts[1])
    except ValueError:
        return None
    if year < 2000 or year > 2200 or month < 1 or month > 12:
        return None
    return year, month


def _parse_iso_date(value: str | None) -> date | None:
    if not value:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError:
        return None


def _parse_bounded_int(raw_value, *, default: int, min_value: int, max_value: int) -> int:
    try:
        value = int(raw_value)
    except (TypeError, ValueError):
        value = default
    return max(min_value, min(value, max_value))


def _to_float(value, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError, InvalidOperation):
        return default


def _pct_change(current: int, previous: int) -> float:
    current_i = int(current or 0)
    previous_i = int(previous or 0)
    if previous_i <= 0:
        return 100.0 if current_i > 0 else 0.0
    return round(((current_i - previous_i) * 100.0 / previous_i), 2)


def _build_public_api_daily_trend(days: int = 7) -> list[dict[str, Any]]:
    days = max(1, min(int(days or 7), 31))
    today = timezone.localdate()
    start_date = today - timedelta(days=days - 1)
    raw_rows = list(
        PublicApiAccessLog.objects.filter(created_at__date__gte=start_date)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(
            total=Count("id"),
            errors=Count("id", filter=Q(status_code__gte=400)),
        )
        .order_by("day")
    )
    by_day = {row["day"]: row for row in raw_rows}
    trend = []
    for index in range(days):
        day = start_date + timedelta(days=index)
        row = by_day.get(day, {})
        total = int(row.get("total") or 0)
        errors = int(row.get("errors") or 0)
        trend.append(
            {
                "day": day,
                "requests": total,
                "errors": errors,
                "error_rate_pct": round((errors * 100.0 / total), 2) if total else 0.0,
            }
        )
    return trend


def _preview_deactivate_idle_api_clients(idle_days: int, limit: int) -> dict[str, Any]:
    idle_days = max(1, min(int(idle_days or 30), 365))
    limit = max(1, min(int(limit or 100), 500))
    cutoff = timezone.now() - timedelta(days=idle_days)
    recent_client_ids = set(
        PublicApiAccessLog.objects.filter(created_at__gte=cutoff)
        .values_list("client_id", flat=True)
        .distinct()
    )
    candidates = list(
        PublicApiClient.objects.filter(activo=True)
        .exclude(id__in=recent_client_ids)
        .order_by("id")
        .values("id", "nombre")[:limit]
    )
    return {
        "idle_days": idle_days,
        "limit": limit,
        "candidates": len(candidates),
        "deactivated": 0,
        "cutoff": cutoff.isoformat(),
        "candidate_clients": candidates,
        "dry_run": True,
    }


def _preview_purge_api_logs(retain_days: int, max_delete: int) -> dict[str, Any]:
    retain_days = max(1, min(int(retain_days or 90), 3650))
    max_delete = max(1, min(int(max_delete or 5000), 50000))
    cutoff = timezone.now() - timedelta(days=retain_days)
    total_candidates = PublicApiAccessLog.objects.filter(created_at__lt=cutoff).count()
    return {
        "retain_days": retain_days,
        "max_delete": max_delete,
        "cutoff": cutoff.isoformat(),
        "candidates": int(total_candidates),
        "deleted": 0,
        "remaining_candidates": int(total_candidates),
        "would_delete": min(int(total_candidates), max_delete),
        "dry_run": True,
    }


class ApiTokenAuthView(ObtainAuthToken):
    permission_classes = [AllowAny]
    authentication_classes = []

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data["user"]
        token, _ = Token.objects.get_or_create(user=user)
        return Response(
            {
                "token": token.key,
                "user": {
                    "id": user.id,
                    "username": user.get_username(),
                    "is_superuser": bool(user.is_superuser),
                },
            },
            status=status.HTTP_200_OK,
        )


class ApiTokenRotateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        Token.objects.filter(user=request.user).delete()
        token = Token.objects.create(user=request.user)
        return Response({"token": token.key}, status=status.HTTP_200_OK)


class ApiTokenRevokeView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        deleted, _ = Token.objects.filter(user=request.user).delete()
        return Response({"revoked": bool(deleted)}, status=status.HTTP_200_OK)


class ApiAuthMeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        group_names = list(user.groups.values_list("name", flat=True)) if user.is_authenticated else []
        return Response(
            {
                "id": user.id,
                "username": user.get_username(),
                "email": getattr(user, "email", "") or "",
                "is_superuser": bool(user.is_superuser),
                "is_staff": bool(user.is_staff),
                "groups": sorted(group_names),
                "permissions": {
                    "can_view_compras": can_view_compras(user),
                    "can_manage_compras": can_manage_compras(user),
                    "can_view_inventario": can_view_inventario(user),
                    "can_manage_inventario": can_manage_inventario(user),
                    "can_view_maestros": can_view_maestros(user),
                    "can_view_audit": can_view_audit(user),
                },
            },
            status=status.HTTP_200_OK,
        )


class AuditLogListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not can_view_audit(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar bitácora."},
                status=status.HTTP_403_FORBIDDEN,
            )

        q = (request.GET.get("q") or "").strip()
        action = (request.GET.get("action") or "").strip().upper()
        model_name = (request.GET.get("model") or "").strip()
        user_id_raw = (request.GET.get("user_id") or "").strip()
        limit = _parse_bounded_int(request.GET.get("limit", 200), default=200, min_value=1, max_value=1000)
        offset = _parse_bounded_int(request.GET.get("offset", 0), default=0, min_value=0, max_value=200000)
        sort_by = (request.GET.get("sort_by") or "timestamp").strip().lower()
        sort_dir = (request.GET.get("sort_dir") or "desc").strip().lower()
        allowed_sort = {
            "timestamp": "timestamp",
            "action": "action",
            "model": "model",
            "object_id": "object_id",
            "user": "user__username",
            "id": "id",
        }
        if sort_by not in allowed_sort:
            return Response(
                {"detail": "sort_by inválido. Usa timestamp, action, model, object_id, user o id."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if sort_dir not in {"asc", "desc"}:
            return Response(
                {"detail": "sort_dir inválido. Usa asc o desc."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = AuditLog.objects.select_related("user").order_by("-timestamp", "-id")
        if q:
            qs = qs.filter(
                Q(action__icontains=q)
                | Q(model__icontains=q)
                | Q(object_id__icontains=q)
                | Q(payload__icontains=q)
                | Q(user__username__icontains=q)
            )
        if action:
            qs = qs.filter(action=action)
        if model_name:
            qs = qs.filter(model__icontains=model_name)
        if user_id_raw:
            try:
                user_id = int(user_id_raw)
            except ValueError:
                return Response({"detail": "user_id inválido."}, status=status.HTTP_400_BAD_REQUEST)
            qs = qs.filter(user_id=user_id)
        sort_field = allowed_sort[sort_by]
        order_expr = sort_field if sort_dir == "asc" else f"-{sort_field}"
        qs = qs.order_by(order_expr, "-id")

        total = qs.count()
        rows = []
        for log in qs[offset : offset + limit]:
            rows.append(
                {
                    "id": log.id,
                    "timestamp": log.timestamp,
                    "action": log.action,
                    "model": log.model,
                    "object_id": log.object_id,
                    "user_id": log.user_id,
                    "user": log.user.username if log.user_id else "",
                    "payload": log.payload or {},
                }
            )
        return Response(
            {
                "filters": {
                    "q": q,
                    "action": action,
                    "model": model_name,
                    "user_id": user_id_raw,
                    "limit": limit,
                    "offset": offset,
                    "sort_by": sort_by,
                    "sort_dir": sort_dir,
                },
                "totales": {
                    "rows": total,
                    "rows_total": total,
                    "rows_returned": len(rows),
                    "returned": len(rows),
                },
                "items": rows,
            },
            status=status.HTTP_200_OK,
        )


def _can_approve_ajustes(user) -> bool:
    return has_any_role(user, ROLE_ADMIN, ROLE_DG)


def _serialize_ajuste_row(ajuste: AjusteInventario) -> dict:
    return {
        "id": ajuste.id,
        "folio": ajuste.folio,
        "insumo_id": ajuste.insumo_id,
        "insumo": ajuste.insumo.nombre if ajuste.insumo_id else "",
        "cantidad_sistema": str(ajuste.cantidad_sistema),
        "cantidad_fisica": str(ajuste.cantidad_fisica),
        "delta": str(_to_decimal(ajuste.cantidad_fisica) - _to_decimal(ajuste.cantidad_sistema)),
        "motivo": ajuste.motivo,
        "estatus": ajuste.estatus,
        "solicitado_por": ajuste.solicitado_por.username if ajuste.solicitado_por_id else "",
        "aprobado_por": ajuste.aprobado_por.username if ajuste.aprobado_por_id else "",
        "comentario_revision": ajuste.comentario_revision or "",
        "creado_en": ajuste.creado_en,
        "aprobado_en": ajuste.aprobado_en,
        "aplicado_en": ajuste.aplicado_en,
    }


def _serialize_forecast_compare(compare: dict | None, *, top: int = 120) -> dict:
    if not compare:
        return {
            "target_start": "",
            "target_end": "",
            "sucursal_id": None,
            "sucursal_nombre": "",
            "escenario": "base",
            "rows": [],
            "totals": {
                "forecast_total": 0.0,
                "solicitud_total": 0.0,
                "delta_total": 0.0,
                "ok_count": 0,
                "sobre_count": 0,
                "bajo_count": 0,
                "sin_base_count": 0,
                "en_rango_count": 0,
                "sobre_rango_count": 0,
                "bajo_rango_count": 0,
            },
        }

    rows = []
    for row in (compare.get("rows") or [])[:top]:
        variacion = row.get("variacion_pct")
        rows.append(
            {
                "receta_id": int(row.get("receta_id") or 0),
                "receta": row.get("receta") or "",
                "forecast_qty": _to_float(row.get("forecast_qty")),
                "forecast_base": _to_float(row.get("forecast_base")),
                "forecast_low": _to_float(row.get("forecast_low")),
                "forecast_high": _to_float(row.get("forecast_high")),
                "solicitud_qty": _to_float(row.get("solicitud_qty")),
                "delta_qty": _to_float(row.get("delta_qty")),
                "variacion_pct": _to_float(variacion) if variacion is not None else None,
                "status": row.get("status") or "",
                "status_rango": row.get("status_rango") or "",
            }
        )
    totals = compare.get("totals") or {}
    return {
        "target_start": str(compare.get("target_start") or ""),
        "target_end": str(compare.get("target_end") or ""),
        "sucursal_id": compare.get("sucursal_id"),
        "sucursal_nombre": compare.get("sucursal_nombre") or "",
        "escenario": compare.get("escenario") or "base",
        "rows": rows,
        "totals": {
            "forecast_total": _to_float(totals.get("forecast_total")),
            "solicitud_total": _to_float(totals.get("solicitud_total")),
            "delta_total": _to_float(totals.get("delta_total")),
            "ok_count": int(totals.get("ok_count") or 0),
            "sobre_count": int(totals.get("sobre_count") or 0),
            "bajo_count": int(totals.get("bajo_count") or 0),
            "sin_base_count": int(totals.get("sin_base_count") or 0),
            "en_rango_count": int(totals.get("en_rango_count") or 0),
            "sobre_rango_count": int(totals.get("sobre_rango_count") or 0),
            "bajo_rango_count": int(totals.get("bajo_rango_count") or 0),
        },
    }


def _ventas_pipeline_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    scope = payload.get("scope") or {}
    totals = payload.get("totales") or {}
    by_alcance = payload.get("solicitud_by_alcance") or {}
    by_sucursal = payload.get("by_sucursal") or []
    by_sucursal_status = totals.get("by_sucursal_status") or {}
    rows_status = totals.get("rows_status") or {}
    rows = payload.get("rows") or []
    periodo = str(scope.get("periodo") or _normalize_periodo_mes(None)).replace("-", "")
    filename = f"ventas_pipeline_{periodo}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Periodo", scope.get("periodo") or ""])
        ws_resumen.append(["Sucursal", scope.get("sucursal") or "Todas"])
        ws_resumen.append(["Incluir preparaciones", "SI" if scope.get("incluir_preparaciones") else "NO"])
        ws_resumen.append(["Filtro q", scope.get("q") or ""])
        ws_resumen.append(["Top", int(scope.get("top") or 0)])
        ws_resumen.append(["Offset", int(scope.get("offset") or 0)])
        ws_resumen.append(["Top sucursales", int(scope.get("top_sucursales") or 0)])
        ws_resumen.append(["Offset sucursales", int(scope.get("offset_sucursales") or 0)])
        ws_resumen.append(["Sort rows", str(scope.get("sort_by") or "delta_abs").upper()])
        ws_resumen.append(["Sort rows dir", str(scope.get("sort_dir") or "desc").upper()])
        ws_resumen.append(["Sort sucursales", str(scope.get("sort_sucursales_by") or "delta_abs").upper()])
        ws_resumen.append(["Sort sucursales dir", str(scope.get("sort_sucursales_dir") or "desc").upper()])
        ws_resumen.append(["Rows filtered", int(totals.get("rows_filtered") or 0)])
        ws_resumen.append(["Rows returned", int(totals.get("rows_returned") or 0)])
        ws_resumen.append(["Sucursales filtered", int(totals.get("by_sucursal_filtered") or 0)])
        ws_resumen.append(["Sucursales returned", int(totals.get("by_sucursal_returned") or 0)])
        ws_resumen.append(["Historial total", float(totals.get("historial_qty") or 0)])
        ws_resumen.append(["Pronostico total", float(totals.get("pronostico_qty") or 0)])
        ws_resumen.append(["Solicitud total", float(totals.get("solicitud_qty") or 0)])
        ws_resumen.append(["Delta solicitud vs pronostico", float(totals.get("delta_solicitud_vs_pronostico") or 0)])
        ws_resumen.append(["Delta historial vs solicitud", float(totals.get("delta_historial_vs_solicitud") or 0)])
        ws_resumen.append(["Cobertura solicitud %", float(totals.get("cobertura_solicitud_pct")) if totals.get("cobertura_solicitud_pct") is not None else None])
        ws_resumen.append(["Cumplimiento historial %", float(totals.get("cumplimiento_historial_pct")) if totals.get("cumplimiento_historial_pct") is not None else None])
        ws_resumen.append(["Solicitudes MES", float(by_alcance.get("MES") or 0)])
        ws_resumen.append(["Solicitudes SEMANA", float(by_alcance.get("SEMANA") or 0)])
        ws_resumen.append(["Solicitudes FIN_SEMANA", float(by_alcance.get("FIN_SEMANA") or 0)])
        ws_resumen.append(["Rows status SOBRE", int(rows_status.get("SOBRE") or 0)])
        ws_resumen.append(["Rows status BAJO", int(rows_status.get("BAJO") or 0)])
        ws_resumen.append(["Rows status OK", int(rows_status.get("OK") or 0)])
        ws_resumen.append(["Rows status SIN_SOLICITUD", int(rows_status.get("SIN_SOLICITUD") or 0)])
        ws_resumen.append(["Rows status SIN_MOV", int(rows_status.get("SIN_MOV") or 0)])
        ws_resumen.append(["Sucursales status SOBRE", int(by_sucursal_status.get("SOBRE") or 0)])
        ws_resumen.append(["Sucursales status BAJO", int(by_sucursal_status.get("BAJO") or 0)])
        ws_resumen.append(["Sucursales status OK", int(by_sucursal_status.get("OK") or 0)])
        ws_resumen.append(["Sucursales status SIN_SOLICITUD", int(by_sucursal_status.get("SIN_SOLICITUD") or 0)])
        ws_resumen.append(["Sucursales status SIN_MOV", int(by_sucursal_status.get("SIN_MOV") or 0)])

        ws_sucursal = wb.create_sheet("BySucursal")
        ws_sucursal.append(
            [
                "Sucursal ID",
                "Sucursal codigo",
                "Sucursal",
                "Historial",
                "Solicitud",
                "Delta Historial vs Solicitud",
                "Cumplimiento %",
                "Status",
            ]
        )
        for row in by_sucursal:
            ws_sucursal.append(
                [
                    int(row.get("sucursal_id") or 0),
                    row.get("sucursal_codigo") or "",
                    row.get("sucursal") or "",
                    float(row.get("historial_qty") or 0),
                    float(row.get("solicitud_qty") or 0),
                    float(row.get("delta_historial_vs_solicitud") or 0),
                    float(row.get("cumplimiento_pct")) if row.get("cumplimiento_pct") is not None else None,
                    row.get("status") or "",
                ]
            )

        ws_detalle = wb.create_sheet("Detalle")
        ws_detalle.append(
            [
                "Receta ID",
                "Receta",
                "Historial",
                "Pronostico",
                "Solicitud",
                "Delta Solicitud vs Pronostico",
                "Delta Historial vs Solicitud",
                "Cobertura %",
                "Cumplimiento %",
                "Status",
            ]
        )
        for row in rows:
            ws_detalle.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    float(row.get("historial_qty") or 0),
                    float(row.get("pronostico_qty") or 0),
                    float(row.get("solicitud_qty") or 0),
                    float(row.get("delta_solicitud_vs_pronostico") or 0),
                    float(row.get("delta_historial_vs_solicitud") or 0),
                    float(row.get("cobertura_pct")) if row.get("cobertura_pct") is not None else None,
                    float(row.get("cumplimiento_pct")) if row.get("cumplimiento_pct") is not None else None,
                    row.get("status") or "",
                ]
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Periodo", scope.get("periodo") or ""])
    writer.writerow(["Sucursal", scope.get("sucursal") or "Todas"])
    writer.writerow(["Incluir preparaciones", "SI" if scope.get("incluir_preparaciones") else "NO"])
    writer.writerow(["Filtro q", scope.get("q") or ""])
    writer.writerow(["Top", int(scope.get("top") or 0)])
    writer.writerow(["Offset", int(scope.get("offset") or 0)])
    writer.writerow(["Top sucursales", int(scope.get("top_sucursales") or 0)])
    writer.writerow(["Offset sucursales", int(scope.get("offset_sucursales") or 0)])
    writer.writerow(["Sort rows", str(scope.get("sort_by") or "delta_abs").upper()])
    writer.writerow(["Sort rows dir", str(scope.get("sort_dir") or "desc").upper()])
    writer.writerow(["Sort sucursales", str(scope.get("sort_sucursales_by") or "delta_abs").upper()])
    writer.writerow(["Sort sucursales dir", str(scope.get("sort_sucursales_dir") or "desc").upper()])
    writer.writerow(["Rows filtered", int(totals.get("rows_filtered") or 0)])
    writer.writerow(["Rows returned", int(totals.get("rows_returned") or 0)])
    writer.writerow(["Sucursales filtered", int(totals.get("by_sucursal_filtered") or 0)])
    writer.writerow(["Sucursales returned", int(totals.get("by_sucursal_returned") or 0)])
    writer.writerow(["Historial total", f"{Decimal(str(totals.get('historial_qty') or 0)):.3f}"])
    writer.writerow(["Pronostico total", f"{Decimal(str(totals.get('pronostico_qty') or 0)):.3f}"])
    writer.writerow(["Solicitud total", f"{Decimal(str(totals.get('solicitud_qty') or 0)):.3f}"])
    writer.writerow(["Delta solicitud vs pronostico", f"{Decimal(str(totals.get('delta_solicitud_vs_pronostico') or 0)):.3f}"])
    writer.writerow(["Delta historial vs solicitud", f"{Decimal(str(totals.get('delta_historial_vs_solicitud') or 0)):.3f}"])
    writer.writerow(
        [
            "Cobertura solicitud %",
            f"{Decimal(str(totals.get('cobertura_solicitud_pct'))):.1f}" if totals.get("cobertura_solicitud_pct") is not None else "",
        ]
    )
    writer.writerow(
        [
            "Cumplimiento historial %",
            f"{Decimal(str(totals.get('cumplimiento_historial_pct'))):.1f}" if totals.get("cumplimiento_historial_pct") is not None else "",
        ]
    )
    writer.writerow(["Solicitudes MES", f"{Decimal(str(by_alcance.get('MES') or 0)):.3f}"])
    writer.writerow(["Solicitudes SEMANA", f"{Decimal(str(by_alcance.get('SEMANA') or 0)):.3f}"])
    writer.writerow(["Solicitudes FIN_SEMANA", f"{Decimal(str(by_alcance.get('FIN_SEMANA') or 0)):.3f}"])
    writer.writerow(["Rows status SOBRE", int(rows_status.get("SOBRE") or 0)])
    writer.writerow(["Rows status BAJO", int(rows_status.get("BAJO") or 0)])
    writer.writerow(["Rows status OK", int(rows_status.get("OK") or 0)])
    writer.writerow(["Rows status SIN_SOLICITUD", int(rows_status.get("SIN_SOLICITUD") or 0)])
    writer.writerow(["Rows status SIN_MOV", int(rows_status.get("SIN_MOV") or 0)])
    writer.writerow(["Sucursales status SOBRE", int(by_sucursal_status.get("SOBRE") or 0)])
    writer.writerow(["Sucursales status BAJO", int(by_sucursal_status.get("BAJO") or 0)])
    writer.writerow(["Sucursales status OK", int(by_sucursal_status.get("OK") or 0)])
    writer.writerow(["Sucursales status SIN_SOLICITUD", int(by_sucursal_status.get("SIN_SOLICITUD") or 0)])
    writer.writerow(["Sucursales status SIN_MOV", int(by_sucursal_status.get("SIN_MOV") or 0)])
    writer.writerow([])
    writer.writerow(["BY_SUCURSAL"])
    writer.writerow(
        [
            "sucursal_id",
            "sucursal_codigo",
            "sucursal",
            "historial_qty",
            "solicitud_qty",
            "delta_historial_vs_solicitud",
            "cumplimiento_pct",
            "status",
        ]
    )
    for row in by_sucursal:
        writer.writerow(
            [
                int(row.get("sucursal_id") or 0),
                row.get("sucursal_codigo") or "",
                row.get("sucursal") or "",
                f"{Decimal(str(row.get('historial_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('solicitud_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('delta_historial_vs_solicitud') or 0)):.3f}",
                f"{Decimal(str(row.get('cumplimiento_pct'))):.1f}" if row.get("cumplimiento_pct") is not None else "",
                row.get("status") or "",
            ]
        )
    writer.writerow([])
    writer.writerow(
        [
            "receta_id",
            "receta",
            "historial_qty",
            "pronostico_qty",
            "solicitud_qty",
            "delta_solicitud_vs_pronostico",
            "delta_historial_vs_solicitud",
            "cobertura_pct",
            "cumplimiento_pct",
            "status",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                f"{Decimal(str(row.get('historial_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('pronostico_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('solicitud_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('delta_solicitud_vs_pronostico') or 0)):.3f}",
                f"{Decimal(str(row.get('delta_historial_vs_solicitud') or 0)):.3f}",
                f"{Decimal(str(row.get('cobertura_pct'))):.1f}" if row.get("cobertura_pct") is not None else "",
                f"{Decimal(str(row.get('cumplimiento_pct'))):.1f}" if row.get("cumplimiento_pct") is not None else "",
                row.get("status") or "",
            ]
        )
    return response


def _forecast_estadistico_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    scope = payload.get("scope") or {}
    totals = payload.get("totals") or {}
    rows = payload.get("rows") or []
    compare = payload.get("compare_solicitud") or {}
    include_compare = bool(compare.get("rows"))
    start_txt = str(scope.get("target_start") or "").replace("-", "")
    end_txt = str(scope.get("target_end") or "").replace("-", "")
    if not start_txt or not end_txt:
        today_txt = timezone.localdate().strftime("%Y%m%d")
        start_txt = start_txt or today_txt
        end_txt = end_txt or today_txt
    filename = f"api_forecast_estadistico_{start_txt}_{end_txt}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Alcance", str(scope.get("alcance") or "").upper()])
        ws_resumen.append(["Periodo", scope.get("periodo") or ""])
        ws_resumen.append(["Rango inicio", scope.get("target_start") or ""])
        ws_resumen.append(["Rango fin", scope.get("target_end") or ""])
        ws_resumen.append(["Sucursal", scope.get("sucursal_nombre") or "Todas"])
        ws_resumen.append(["Escenario compare", str(scope.get("escenario_compare") or "base").upper()])
        ws_resumen.append(["Confianza minima %", float(scope.get("min_confianza_pct") or 0)])
        ws_resumen.append(["Forecast total", float(totals.get("forecast_total") or 0)])
        ws_resumen.append(["Banda baja total", float(totals.get("forecast_low_total") or 0)])
        ws_resumen.append(["Banda alta total", float(totals.get("forecast_high_total") or 0)])
        ws_resumen.append(["Pronostico total", float(totals.get("pronostico_total") or 0)])
        ws_resumen.append(["Delta total", float(totals.get("delta_total") or 0)])

        ws_forecast = wb.create_sheet("Forecast")
        ws_forecast.append(
            [
                "Receta ID",
                "Receta",
                "Forecast",
                "Banda baja",
                "Banda alta",
                "Pronostico actual",
                "Delta",
                "Recomendacion",
                "Confianza %",
                "Desviacion",
                "Muestras",
                "Observaciones",
            ]
        )
        for row in rows:
            ws_forecast.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    float(row.get("forecast_qty") or 0),
                    float(row.get("forecast_low") or 0),
                    float(row.get("forecast_high") or 0),
                    float(row.get("pronostico_actual") or 0),
                    float(row.get("delta") or 0),
                    row.get("recomendacion") or "",
                    float(row.get("confianza") or 0),
                    float(row.get("desviacion") or 0),
                    int(row.get("muestras") or 0),
                    row.get("observaciones") or "",
                ]
            )

        if include_compare:
            ws_compare = wb.create_sheet("CompareSolicitud")
            ws_compare.append(
                [
                    "Receta ID",
                    "Receta",
                    "Forecast",
                    "Forecast base",
                    "Forecast baja",
                    "Forecast alta",
                    "Solicitud",
                    "Delta",
                    "Variacion %",
                    "Status",
                    "Status rango",
                ]
            )
            for row in compare.get("rows") or []:
                ws_compare.append(
                    [
                        int(row.get("receta_id") or 0),
                        row.get("receta") or "",
                        float(row.get("forecast_qty") or 0),
                        float(row.get("forecast_base") or 0),
                        float(row.get("forecast_low") or 0),
                        float(row.get("forecast_high") or 0),
                        float(row.get("solicitud_qty") or 0),
                        float(row.get("delta_qty") or 0),
                        float(row.get("variacion_pct")) if row.get("variacion_pct") is not None else None,
                        row.get("status") or "",
                        row.get("status_rango") or "",
                    ]
                )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Alcance", str(scope.get("alcance") or "").upper()])
    writer.writerow(["Periodo", scope.get("periodo") or ""])
    writer.writerow(["Rango inicio", scope.get("target_start") or ""])
    writer.writerow(["Rango fin", scope.get("target_end") or ""])
    writer.writerow(["Sucursal", scope.get("sucursal_nombre") or "Todas"])
    writer.writerow(["Escenario compare", str(scope.get("escenario_compare") or "base").upper()])
    writer.writerow(["Confianza minima %", f"{Decimal(str(scope.get('min_confianza_pct') or 0)):.1f}"])
    writer.writerow(["Forecast total", f"{Decimal(str(totals.get('forecast_total') or 0)):.3f}"])
    writer.writerow(["Banda baja total", f"{Decimal(str(totals.get('forecast_low_total') or 0)):.3f}"])
    writer.writerow(["Banda alta total", f"{Decimal(str(totals.get('forecast_high_total') or 0)):.3f}"])
    writer.writerow(["Pronostico total", f"{Decimal(str(totals.get('pronostico_total') or 0)):.3f}"])
    writer.writerow(["Delta total", f"{Decimal(str(totals.get('delta_total') or 0)):.3f}"])
    writer.writerow([])
    writer.writerow(
        [
            "receta_id",
            "receta",
            "forecast",
            "banda_baja",
            "banda_alta",
            "pronostico_actual",
            "delta",
            "recomendacion",
            "confianza_pct",
            "desviacion",
            "muestras",
            "observaciones",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                f"{Decimal(str(row.get('forecast_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_low') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_high') or 0)):.3f}",
                f"{Decimal(str(row.get('pronostico_actual') or 0)):.3f}",
                f"{Decimal(str(row.get('delta') or 0)):.3f}",
                row.get("recomendacion") or "",
                f"{Decimal(str(row.get('confianza') or 0)):.1f}",
                f"{Decimal(str(row.get('desviacion') or 0)):.3f}",
                int(row.get("muestras") or 0),
                row.get("observaciones") or "",
            ]
        )
    if include_compare:
        writer.writerow([])
        writer.writerow(["COMPARE_SOLICITUD"])
        writer.writerow(
            [
                "receta_id",
                "receta",
                "forecast",
                "forecast_base",
                "forecast_baja",
                "forecast_alta",
                "solicitud",
                "delta",
                "variacion_pct",
                "status",
                "status_rango",
            ]
        )
        for row in compare.get("rows") or []:
            writer.writerow(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    f"{Decimal(str(row.get('forecast_qty') or 0)):.3f}",
                    f"{Decimal(str(row.get('forecast_base') or 0)):.3f}",
                    f"{Decimal(str(row.get('forecast_low') or 0)):.3f}",
                    f"{Decimal(str(row.get('forecast_high') or 0)):.3f}",
                    f"{Decimal(str(row.get('solicitud_qty') or 0)):.3f}",
                    f"{Decimal(str(row.get('delta_qty') or 0)):.3f}",
                    f"{Decimal(str(row.get('variacion_pct'))):.1f}" if row.get("variacion_pct") is not None else "",
                    row.get("status") or "",
                    row.get("status_rango") or "",
                ]
            )
    return response


def _forecast_estadistico_guardar_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    scope = payload.get("scope") or {}
    totals = payload.get("totals") or {}
    persisted = payload.get("persisted") or {}
    rows = payload.get("rows") or []
    applied_rows = payload.get("applied_rows") or []
    start_txt = str(scope.get("target_start") or "").replace("-", "")
    end_txt = str(scope.get("target_end") or "").replace("-", "")
    if not start_txt or not end_txt:
        today_txt = timezone.localdate().strftime("%Y%m%d")
        start_txt = start_txt or today_txt
        end_txt = end_txt or today_txt
    filename = f"api_forecast_estadistico_guardar_{start_txt}_{end_txt}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Alcance", str(scope.get("alcance") or "").upper()])
        ws_resumen.append(["Periodo", scope.get("periodo") or ""])
        ws_resumen.append(["Rango inicio", scope.get("target_start") or ""])
        ws_resumen.append(["Rango fin", scope.get("target_end") or ""])
        ws_resumen.append(["Sucursal", scope.get("sucursal_nombre") or "Todas"])
        ws_resumen.append(["Escenario", str(scope.get("escenario") or "base").upper()])
        ws_resumen.append(["Confianza minima %", float(scope.get("min_confianza_pct") or 0)])
        ws_resumen.append(["Filtradas por confianza", int(scope.get("filtered_conf") or 0)])
        ws_resumen.append(["Forecast total", float(totals.get("forecast_total") or 0)])
        ws_resumen.append(["Banda baja total", float(totals.get("forecast_low_total") or 0)])
        ws_resumen.append(["Banda alta total", float(totals.get("forecast_high_total") or 0)])
        ws_resumen.append(["Pronostico total", float(totals.get("pronostico_total") or 0)])
        ws_resumen.append(["Delta total", float(totals.get("delta_total") or 0)])
        ws_resumen.append(["Created", int(persisted.get("created") or 0)])
        ws_resumen.append(["Updated", int(persisted.get("updated") or 0)])
        ws_resumen.append(["Skipped existing", int(persisted.get("skipped_existing") or 0)])
        ws_resumen.append(["Skipped invalid", int(persisted.get("skipped_invalid") or 0)])
        ws_resumen.append(["Applied", int(persisted.get("applied") or 0)])

        ws_forecast = wb.create_sheet("Forecast")
        ws_forecast.append(
            [
                "Receta ID",
                "Receta",
                "Forecast",
                "Banda baja",
                "Banda alta",
                "Pronostico actual",
                "Delta",
                "Recomendacion",
                "Confianza %",
                "Desviacion",
                "Muestras",
                "Observaciones",
            ]
        )
        for row in rows:
            ws_forecast.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    float(row.get("forecast_qty") or 0),
                    float(row.get("forecast_low") or 0),
                    float(row.get("forecast_high") or 0),
                    float(row.get("pronostico_actual") or 0),
                    float(row.get("delta") or 0),
                    row.get("recomendacion") or "",
                    float(row.get("confianza") or 0),
                    float(row.get("desviacion") or 0),
                    int(row.get("muestras") or 0),
                    row.get("observaciones") or "",
                ]
            )

        ws_applied = wb.create_sheet("Applied")
        ws_applied.append(["Receta ID", "Receta", "Escenario", "Cantidad anterior", "Cantidad nueva", "Accion"])
        for row in applied_rows:
            ws_applied.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    row.get("escenario") or "",
                    float(row.get("cantidad_anterior") or 0),
                    float(row.get("cantidad_nueva") or 0),
                    row.get("accion") or "",
                ]
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Alcance", str(scope.get("alcance") or "").upper()])
    writer.writerow(["Periodo", scope.get("periodo") or ""])
    writer.writerow(["Rango inicio", scope.get("target_start") or ""])
    writer.writerow(["Rango fin", scope.get("target_end") or ""])
    writer.writerow(["Sucursal", scope.get("sucursal_nombre") or "Todas"])
    writer.writerow(["Escenario", str(scope.get("escenario") or "base").upper()])
    writer.writerow(["Confianza minima %", f"{Decimal(str(scope.get('min_confianza_pct') or 0)):.1f}"])
    writer.writerow(["Filtradas por confianza", int(scope.get("filtered_conf") or 0)])
    writer.writerow(["Forecast total", f"{Decimal(str(totals.get('forecast_total') or 0)):.3f}"])
    writer.writerow(["Banda baja total", f"{Decimal(str(totals.get('forecast_low_total') or 0)):.3f}"])
    writer.writerow(["Banda alta total", f"{Decimal(str(totals.get('forecast_high_total') or 0)):.3f}"])
    writer.writerow(["Pronostico total", f"{Decimal(str(totals.get('pronostico_total') or 0)):.3f}"])
    writer.writerow(["Delta total", f"{Decimal(str(totals.get('delta_total') or 0)):.3f}"])
    writer.writerow(["Created", int(persisted.get("created") or 0)])
    writer.writerow(["Updated", int(persisted.get("updated") or 0)])
    writer.writerow(["Skipped existing", int(persisted.get("skipped_existing") or 0)])
    writer.writerow(["Skipped invalid", int(persisted.get("skipped_invalid") or 0)])
    writer.writerow(["Applied", int(persisted.get("applied") or 0)])
    writer.writerow([])
    writer.writerow(["FORECAST"])
    writer.writerow(
        [
            "receta_id",
            "receta",
            "forecast",
            "banda_baja",
            "banda_alta",
            "pronostico_actual",
            "delta",
            "recomendacion",
            "confianza_pct",
            "desviacion",
            "muestras",
            "observaciones",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                f"{Decimal(str(row.get('forecast_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_low') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_high') or 0)):.3f}",
                f"{Decimal(str(row.get('pronostico_actual') or 0)):.3f}",
                f"{Decimal(str(row.get('delta') or 0)):.3f}",
                row.get("recomendacion") or "",
                f"{Decimal(str(row.get('confianza') or 0)):.1f}",
                f"{Decimal(str(row.get('desviacion') or 0)):.3f}",
                int(row.get("muestras") or 0),
                row.get("observaciones") or "",
            ]
        )
    writer.writerow([])
    writer.writerow(["APPLIED"])
    writer.writerow(["receta_id", "receta", "escenario", "cantidad_anterior", "cantidad_nueva", "accion"])
    for row in applied_rows:
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                row.get("escenario") or "",
                f"{Decimal(str(row.get('cantidad_anterior') or 0)):.3f}",
                f"{Decimal(str(row.get('cantidad_nueva') or 0)):.3f}",
                row.get("accion") or "",
            ]
        )
    return response


def _forecast_backtest_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    scope = payload.get("scope") or {}
    totals = payload.get("totals") or {}
    windows = payload.get("windows") or []
    alcance = str(scope.get("alcance") or "mes").lower()
    fecha_base = str(scope.get("fecha_base") or timezone.localdate().isoformat()).replace("-", "")
    scenario = str(scope.get("escenario") or "base").lower()
    filename = f"api_forecast_backtest_{alcance}_{fecha_base}_{scenario}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Alcance", str(scope.get("alcance") or "").upper()])
        ws_resumen.append(["Fecha base", str(scope.get("fecha_base") or "")])
        ws_resumen.append(["Escenario", str(scope.get("escenario") or "base").upper()])
        ws_resumen.append(["Sucursal", scope.get("sucursal_nombre") or "Todas"])
        ws_resumen.append(["Confianza minima %", float(scope.get("min_confianza_pct") or 0)])
        ws_resumen.append(["Ventanas evaluadas", int(totals.get("windows_evaluated") or 0)])
        ws_resumen.append(["Forecast total", float(totals.get("forecast_total") or 0)])
        ws_resumen.append(["Real total", float(totals.get("actual_total") or 0)])
        ws_resumen.append(["Bias total", float(totals.get("bias_total") or 0)])
        ws_resumen.append(["MAE promedio", float(totals.get("mae_promedio") or 0)])
        ws_resumen.append(["MAPE promedio", float(totals.get("mape_promedio")) if totals.get("mape_promedio") is not None else None])

        ws_windows = wb.create_sheet("Ventanas")
        ws_windows.append(["Inicio", "Fin", "Periodo", "Recetas", "Forecast", "Real", "Bias", "MAE", "MAPE"])
        for w in windows:
            ws_windows.append(
                [
                    w.get("window_start") or "",
                    w.get("window_end") or "",
                    w.get("periodo") or "",
                    int(w.get("recetas_count") or 0),
                    float(w.get("forecast_total") or 0),
                    float(w.get("actual_total") or 0),
                    float(w.get("bias_total") or 0),
                    float(w.get("mae") or 0),
                    float(w.get("mape")) if w.get("mape") is not None else None,
                ]
            )

        ws_top = wb.create_sheet("TopErrores")
        ws_top.append(
            [
                "Inicio",
                "Fin",
                "Periodo",
                "Receta ID",
                "Receta",
                "Forecast",
                "Real",
                "Delta",
                "Abs Error",
                "Variacion %",
                "Status",
            ]
        )
        for w in windows:
            for row in w.get("top_errors") or []:
                ws_top.append(
                    [
                        w.get("window_start") or "",
                        w.get("window_end") or "",
                        w.get("periodo") or "",
                        int(row.get("receta_id") or 0),
                        row.get("receta") or "",
                        float(row.get("forecast_qty") or 0),
                        float(row.get("actual_qty") or 0),
                        float(row.get("delta_qty") or 0),
                        float(row.get("abs_error") or 0),
                        float(row.get("variacion_pct")) if row.get("variacion_pct") is not None else None,
                        row.get("status") or "",
                    ]
                )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Alcance", str(scope.get("alcance") or "").upper()])
    writer.writerow(["Fecha base", str(scope.get("fecha_base") or "")])
    writer.writerow(["Escenario", str(scope.get("escenario") or "base").upper()])
    writer.writerow(["Sucursal", scope.get("sucursal_nombre") or "Todas"])
    writer.writerow(["Confianza minima %", f"{Decimal(str(scope.get('min_confianza_pct') or 0)):.1f}"])
    writer.writerow(["Ventanas evaluadas", int(totals.get("windows_evaluated") or 0)])
    writer.writerow(["Forecast total", f"{Decimal(str(totals.get('forecast_total') or 0)):.3f}"])
    writer.writerow(["Real total", f"{Decimal(str(totals.get('actual_total') or 0)):.3f}"])
    writer.writerow(["Bias total", f"{Decimal(str(totals.get('bias_total') or 0)):.3f}"])
    writer.writerow(["MAE promedio", f"{Decimal(str(totals.get('mae_promedio') or 0)):.3f}"])
    writer.writerow(
        [
            "MAPE promedio",
            f"{Decimal(str(totals.get('mape_promedio'))):.1f}" if totals.get("mape_promedio") is not None else "",
        ]
    )
    writer.writerow([])
    writer.writerow(["VENTANAS"])
    writer.writerow(["inicio", "fin", "periodo", "recetas", "forecast", "real", "bias", "mae", "mape"])
    for w in windows:
        writer.writerow(
            [
                w.get("window_start") or "",
                w.get("window_end") or "",
                w.get("periodo") or "",
                int(w.get("recetas_count") or 0),
                f"{Decimal(str(w.get('forecast_total') or 0)):.3f}",
                f"{Decimal(str(w.get('actual_total') or 0)):.3f}",
                f"{Decimal(str(w.get('bias_total') or 0)):.3f}",
                f"{Decimal(str(w.get('mae') or 0)):.3f}",
                f"{Decimal(str(w.get('mape'))):.1f}" if w.get("mape") is not None else "",
            ]
        )
    writer.writerow([])
    writer.writerow(["TOP_ERRORES"])
    writer.writerow(
        [
            "inicio",
            "fin",
            "periodo",
            "receta_id",
            "receta",
            "forecast",
            "real",
            "delta",
            "abs_error",
            "variacion_pct",
            "status",
        ]
    )
    for w in windows:
        for row in w.get("top_errors") or []:
            writer.writerow(
                [
                    w.get("window_start") or "",
                    w.get("window_end") or "",
                    w.get("periodo") or "",
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    f"{Decimal(str(row.get('forecast_qty') or 0)):.3f}",
                    f"{Decimal(str(row.get('actual_qty') or 0)):.3f}",
                    f"{Decimal(str(row.get('delta_qty') or 0)):.3f}",
                    f"{Decimal(str(row.get('abs_error') or 0)):.3f}",
                    f"{Decimal(str(row.get('variacion_pct'))):.1f}" if row.get("variacion_pct") is not None else "",
                    row.get("status") or "",
                ]
            )
    return response


def _forecast_insights_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    scope = payload.get("scope") or {}
    totals = payload.get("totales") or {}
    seasonality = payload.get("seasonality") or {}
    month_rows = seasonality.get("by_month") or []
    weekday_rows = seasonality.get("by_weekday") or []
    top_rows = payload.get("top_recetas") or []

    start_txt = str(scope.get("fecha_desde") or "").replace("-", "")
    end_txt = str(scope.get("fecha_hasta") or "").replace("-", "")
    if not start_txt or not end_txt:
        today_txt = timezone.localdate().strftime("%Y%m%d")
        start_txt = start_txt or today_txt
        end_txt = end_txt or today_txt
    filename = f"api_forecast_insights_{start_txt}_{end_txt}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Meses", int(scope.get("months") or 0)])
        ws_resumen.append(["Fecha desde", scope.get("fecha_desde") or ""])
        ws_resumen.append(["Fecha hasta", scope.get("fecha_hasta") or ""])
        ws_resumen.append(["Sucursal", scope.get("sucursal") or "Todas"])
        ws_resumen.append(["Receta", scope.get("receta") or "Todas"])
        ws_resumen.append(["Top", int(scope.get("top") or 0)])
        ws_resumen.append(["Offset top", int(scope.get("offset_top") or 0)])
        ws_resumen.append(["Filas", int(totals.get("filas") or 0)])
        ws_resumen.append(["Dias con venta", int(totals.get("dias_con_venta") or 0)])
        ws_resumen.append(["Recetas", int(totals.get("recetas") or 0)])
        ws_resumen.append(["Top recetas total", int(totals.get("top_recetas_total") or 0)])
        ws_resumen.append(["Top recetas returned", int(totals.get("top_recetas_returned") or 0)])
        ws_resumen.append(["Cantidad total", float(totals.get("cantidad_total") or 0)])
        ws_resumen.append(["Promedio diario", float(totals.get("promedio_diario") or 0)])

        ws_mes = wb.create_sheet("EstacionalidadMes")
        ws_mes.append(["Mes", "Etiqueta", "Muestras", "Promedio", "Indice %"])
        for row in month_rows:
            ws_mes.append(
                [
                    int(row.get("month") or 0),
                    row.get("label") or "",
                    int(row.get("samples") or 0),
                    float(row.get("avg_qty") or 0),
                    float(row.get("index_pct") or 0),
                ]
            )

        ws_dia = wb.create_sheet("EstacionalidadDia")
        ws_dia.append(["Dia semana", "Etiqueta", "Muestras", "Promedio", "Indice %"])
        for row in weekday_rows:
            ws_dia.append(
                [
                    int(row.get("weekday") or 0),
                    row.get("label") or "",
                    int(row.get("samples") or 0),
                    float(row.get("avg_qty") or 0),
                    float(row.get("index_pct") or 0),
                ]
            )

        ws_top = wb.create_sheet("TopRecetas")
        ws_top.append(["Receta ID", "Receta", "Cantidad total", "Promedio dia activo", "Dias con venta", "Participacion %"])
        for row in top_rows:
            ws_top.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    float(row.get("cantidad_total") or 0),
                    float(row.get("promedio_dia_activo") or 0),
                    int(row.get("dias_con_venta") or 0),
                    float(row.get("participacion_pct") or 0),
                ]
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Meses", int(scope.get("months") or 0)])
    writer.writerow(["Fecha desde", scope.get("fecha_desde") or ""])
    writer.writerow(["Fecha hasta", scope.get("fecha_hasta") or ""])
    writer.writerow(["Sucursal", scope.get("sucursal") or "Todas"])
    writer.writerow(["Receta", scope.get("receta") or "Todas"])
    writer.writerow(["Top", int(scope.get("top") or 0)])
    writer.writerow(["Offset top", int(scope.get("offset_top") or 0)])
    writer.writerow(["Filas", int(totals.get("filas") or 0)])
    writer.writerow(["Dias con venta", int(totals.get("dias_con_venta") or 0)])
    writer.writerow(["Recetas", int(totals.get("recetas") or 0)])
    writer.writerow(["Top recetas total", int(totals.get("top_recetas_total") or 0)])
    writer.writerow(["Top recetas returned", int(totals.get("top_recetas_returned") or 0)])
    writer.writerow(["Cantidad total", f"{Decimal(str(totals.get('cantidad_total') or 0)):.3f}"])
    writer.writerow(["Promedio diario", f"{Decimal(str(totals.get('promedio_diario') or 0)):.3f}"])

    writer.writerow([])
    writer.writerow(["ESTACIONALIDAD_MES"])
    writer.writerow(["month", "label", "samples", "avg_qty", "index_pct"])
    for row in month_rows:
        writer.writerow(
            [
                int(row.get("month") or 0),
                row.get("label") or "",
                int(row.get("samples") or 0),
                f"{Decimal(str(row.get('avg_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('index_pct') or 0)):.1f}",
            ]
        )

    writer.writerow([])
    writer.writerow(["ESTACIONALIDAD_DIA"])
    writer.writerow(["weekday", "label", "samples", "avg_qty", "index_pct"])
    for row in weekday_rows:
        writer.writerow(
            [
                int(row.get("weekday") or 0),
                row.get("label") or "",
                int(row.get("samples") or 0),
                f"{Decimal(str(row.get('avg_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('index_pct') or 0)):.1f}",
            ]
        )

    writer.writerow([])
    writer.writerow(["TOP_RECETAS"])
    writer.writerow(["receta_id", "receta", "cantidad_total", "promedio_dia_activo", "dias_con_venta", "participacion_pct"])
    for row in top_rows:
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                f"{Decimal(str(row.get('cantidad_total') or 0)):.3f}",
                f"{Decimal(str(row.get('promedio_dia_activo') or 0)):.3f}",
                int(row.get("dias_con_venta") or 0),
                f"{Decimal(str(row.get('participacion_pct') or 0)):.1f}",
            ]
        )
    return response


def _ventas_historial_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    filters = payload.get("filters") or {}
    totals = payload.get("totales") or {}
    items = payload.get("items") or []
    by_sucursal = totals.get("by_sucursal") or []
    stamp = str(filters.get("periodo") or timezone.localdate().strftime("%Y-%m")).replace("-", "")
    filename = f"ventas_historial_{stamp}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Periodo", filters.get("periodo") or ""])
        ws_resumen.append(["Desde", filters.get("fecha_desde") or ""])
        ws_resumen.append(["Hasta", filters.get("fecha_hasta") or ""])
        ws_resumen.append(["Offset", int(filters.get("offset") or 0)])
        ws_resumen.append(["Rows", int(totals.get("rows") or 0)])
        ws_resumen.append(["Rows total", int(totals.get("rows_total") or totals.get("rows") or 0)])
        ws_resumen.append(["Rows returned", int(totals.get("rows_returned") or totals.get("rows") or 0)])
        ws_resumen.append(["Cantidad total", float(totals.get("cantidad_total") or 0)])
        ws_resumen.append(["Tickets total", int(totals.get("tickets_total") or 0)])
        ws_resumen.append(["Monto total", float(totals.get("monto_total") or 0)])

        ws_detalle = wb.create_sheet("Detalle")
        ws_detalle.append(
            [
                "ID",
                "Fecha",
                "Receta ID",
                "Receta",
                "Codigo point",
                "Sucursal ID",
                "Sucursal",
                "Sucursal codigo",
                "Cantidad",
                "Tickets",
                "Monto total",
                "Fuente",
                "Actualizado en",
            ]
        )
        for row in items:
            ws_detalle.append(
                [
                    int(row.get("id") or 0),
                    row.get("fecha") or "",
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    row.get("codigo_point") or "",
                    int(row.get("sucursal_id") or 0) if row.get("sucursal_id") else None,
                    row.get("sucursal") or "",
                    row.get("sucursal_codigo") or "",
                    float(row.get("cantidad") or 0),
                    int(row.get("tickets") or 0),
                    float(row.get("monto_total") or 0),
                    row.get("fuente") or "",
                    str(row.get("actualizado_en") or ""),
                ]
            )

        ws_sucursal = wb.create_sheet("BySucursal")
        ws_sucursal.append(["Sucursal", "Cantidad total"])
        for row in by_sucursal:
            ws_sucursal.append([row.get("sucursal") or "", float(row.get("cantidad_total") or 0)])

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Periodo", filters.get("periodo") or ""])
    writer.writerow(["Desde", filters.get("fecha_desde") or ""])
    writer.writerow(["Hasta", filters.get("fecha_hasta") or ""])
    writer.writerow(["Offset", int(filters.get("offset") or 0)])
    writer.writerow(["Rows", int(totals.get("rows") or 0)])
    writer.writerow(["Rows total", int(totals.get("rows_total") or totals.get("rows") or 0)])
    writer.writerow(["Rows returned", int(totals.get("rows_returned") or totals.get("rows") or 0)])
    writer.writerow(["Cantidad total", f"{Decimal(str(totals.get('cantidad_total') or 0)):.3f}"])
    writer.writerow(["Tickets total", int(totals.get("tickets_total") or 0)])
    writer.writerow(["Monto total", f"{Decimal(str(totals.get('monto_total') or 0)):.3f}"])

    writer.writerow([])
    writer.writerow(["BY_SUCURSAL"])
    writer.writerow(["sucursal", "cantidad_total"])
    for row in by_sucursal:
        writer.writerow(
            [
                row.get("sucursal") or "",
                f"{Decimal(str(row.get('cantidad_total') or 0)):.3f}",
            ]
        )

    writer.writerow([])
    writer.writerow(["DETALLE"])
    writer.writerow(
        [
            "id",
            "fecha",
            "receta_id",
            "receta",
            "codigo_point",
            "sucursal_id",
            "sucursal",
            "sucursal_codigo",
            "cantidad",
            "tickets",
            "monto_total",
            "fuente",
            "actualizado_en",
        ]
    )
    for row in items:
        writer.writerow(
            [
                int(row.get("id") or 0),
                row.get("fecha") or "",
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                row.get("codigo_point") or "",
                int(row.get("sucursal_id") or 0) if row.get("sucursal_id") else "",
                row.get("sucursal") or "",
                row.get("sucursal_codigo") or "",
                f"{Decimal(str(row.get('cantidad') or 0)):.3f}",
                int(row.get("tickets") or 0),
                f"{Decimal(str(row.get('monto_total') or 0)):.3f}",
                row.get("fuente") or "",
                str(row.get("actualizado_en") or ""),
            ]
        )
    return response


def _ventas_pronostico_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    filters = payload.get("filters") or {}
    totals = payload.get("totales") or {}
    items = payload.get("items") or []
    stamp = str(filters.get("periodo") or timezone.localdate().strftime("%Y-%m")).replace("-", "")
    filename = f"ventas_pronostico_{stamp}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Periodo", filters.get("periodo") or ""])
        ws_resumen.append(["Periodo desde", filters.get("periodo_desde") or ""])
        ws_resumen.append(["Periodo hasta", filters.get("periodo_hasta") or ""])
        ws_resumen.append(["Offset", int(filters.get("offset") or 0)])
        ws_resumen.append(["Rows", int(totals.get("rows") or 0)])
        ws_resumen.append(["Rows total", int(totals.get("rows_total") or totals.get("rows") or 0)])
        ws_resumen.append(["Rows returned", int(totals.get("rows_returned") or totals.get("rows") or 0)])
        ws_resumen.append(["Cantidad total", float(totals.get("cantidad_total") or 0)])
        ws_resumen.append(["Periodos", int(totals.get("periodos_count") or 0)])

        ws_detalle = wb.create_sheet("Detalle")
        ws_detalle.append(
            [
                "ID",
                "Receta ID",
                "Receta",
                "Codigo point",
                "Periodo",
                "Cantidad",
                "Fuente",
                "Actualizado en",
            ]
        )
        for row in items:
            ws_detalle.append(
                [
                    int(row.get("id") or 0),
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    row.get("codigo_point") or "",
                    row.get("periodo") or "",
                    float(row.get("cantidad") or 0),
                    row.get("fuente") or "",
                    str(row.get("actualizado_en") or ""),
                ]
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Periodo", filters.get("periodo") or ""])
    writer.writerow(["Periodo desde", filters.get("periodo_desde") or ""])
    writer.writerow(["Periodo hasta", filters.get("periodo_hasta") or ""])
    writer.writerow(["Offset", int(filters.get("offset") or 0)])
    writer.writerow(["Rows", int(totals.get("rows") or 0)])
    writer.writerow(["Rows total", int(totals.get("rows_total") or totals.get("rows") or 0)])
    writer.writerow(["Rows returned", int(totals.get("rows_returned") or totals.get("rows") or 0)])
    writer.writerow(["Cantidad total", f"{Decimal(str(totals.get('cantidad_total') or 0)):.3f}"])
    writer.writerow(["Periodos", int(totals.get("periodos_count") or 0)])
    writer.writerow([])
    writer.writerow(["DETALLE"])
    writer.writerow(["id", "receta_id", "receta", "codigo_point", "periodo", "cantidad", "fuente", "actualizado_en"])
    for row in items:
        writer.writerow(
            [
                int(row.get("id") or 0),
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                row.get("codigo_point") or "",
                row.get("periodo") or "",
                f"{Decimal(str(row.get('cantidad') or 0)):.3f}",
                row.get("fuente") or "",
                str(row.get("actualizado_en") or ""),
            ]
        )
    return response


def _ventas_solicitudes_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    filters = payload.get("filters") or {}
    totals = payload.get("totales") or {}
    by_alcance = totals.get("by_alcance") or {}
    forecast_counts = totals.get("forecast_ref_status") or {}
    items = payload.get("items") or []
    stamp = str(filters.get("periodo") or timezone.localdate().strftime("%Y-%m")).replace("-", "")
    filename = f"ventas_solicitudes_{stamp}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Periodo", filters.get("periodo") or ""])
        ws_resumen.append(["Alcance", filters.get("alcance") or ""])
        ws_resumen.append(["Desde", filters.get("fecha_desde") or ""])
        ws_resumen.append(["Hasta", filters.get("fecha_hasta") or ""])
        ws_resumen.append(["Include forecast ref", "SI" if filters.get("include_forecast_ref") else "NO"])
        ws_resumen.append(["Forecast status", filters.get("forecast_status") or ""])
        ws_resumen.append(["Forecast delta min", float(filters.get("forecast_delta_min") or 0)])
        ws_resumen.append(["Offset", int(filters.get("offset") or 0)])
        ws_resumen.append(["Sort by", filters.get("sort_by") or ""])
        ws_resumen.append(["Sort dir", filters.get("sort_dir") or ""])
        ws_resumen.append(["Rows", int(totals.get("rows") or 0)])
        ws_resumen.append(["Rows total", int(totals.get("rows_total") or totals.get("rows") or 0)])
        ws_resumen.append(["Rows returned", int(totals.get("rows_returned") or totals.get("rows") or 0)])
        ws_resumen.append(["Cantidad total", float(totals.get("cantidad_total") or 0)])
        ws_resumen.append(["MES", int(by_alcance.get(SolicitudVenta.ALCANCE_MES) or 0)])
        ws_resumen.append(["SEMANA", int(by_alcance.get(SolicitudVenta.ALCANCE_SEMANA) or 0)])
        ws_resumen.append(["FIN_SEMANA", int(by_alcance.get(SolicitudVenta.ALCANCE_FIN_SEMANA) or 0)])
        ws_resumen.append(["Forecast status SOBRE", int(forecast_counts.get("SOBRE") or 0)])
        ws_resumen.append(["Forecast status BAJO", int(forecast_counts.get("BAJO") or 0)])
        ws_resumen.append(["Forecast status OK", int(forecast_counts.get("OK") or 0)])
        ws_resumen.append(["Forecast status SIN_FORECAST", int(forecast_counts.get("SIN_FORECAST") or 0)])

        ws_detalle = wb.create_sheet("Detalle")
        ws_detalle.append(
            [
                "ID",
                "Receta ID",
                "Receta",
                "Codigo point",
                "Sucursal ID",
                "Sucursal",
                "Sucursal codigo",
                "Alcance",
                "Periodo",
                "Fecha inicio",
                "Fecha fin",
                "Cantidad",
                "Forecast status",
                "Forecast qty",
                "Delta solicitud vs forecast",
                "Fuente",
                "Actualizado en",
            ]
        )
        for row in items:
            forecast_ref = row.get("forecast_ref") or {}
            ws_detalle.append(
                [
                    int(row.get("id") or 0),
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    row.get("codigo_point") or "",
                    int(row.get("sucursal_id") or 0) if row.get("sucursal_id") else None,
                    row.get("sucursal") or "",
                    row.get("sucursal_codigo") or "",
                    row.get("alcance") or "",
                    row.get("periodo") or "",
                    row.get("fecha_inicio") or "",
                    row.get("fecha_fin") or "",
                    float(row.get("cantidad") or 0),
                    forecast_ref.get("status") or "",
                    float(forecast_ref.get("forecast_qty") or 0),
                    float(forecast_ref.get("delta_solicitud_vs_forecast") or 0),
                    row.get("fuente") or "",
                    str(row.get("actualizado_en") or ""),
                ]
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Periodo", filters.get("periodo") or ""])
    writer.writerow(["Alcance", filters.get("alcance") or ""])
    writer.writerow(["Desde", filters.get("fecha_desde") or ""])
    writer.writerow(["Hasta", filters.get("fecha_hasta") or ""])
    writer.writerow(["Include forecast ref", "SI" if filters.get("include_forecast_ref") else "NO"])
    writer.writerow(["Forecast status", filters.get("forecast_status") or ""])
    writer.writerow(["Forecast delta min", f"{Decimal(str(filters.get('forecast_delta_min') or 0)):.3f}"])
    writer.writerow(["Offset", int(filters.get("offset") or 0)])
    writer.writerow(["Sort by", filters.get("sort_by") or ""])
    writer.writerow(["Sort dir", filters.get("sort_dir") or ""])
    writer.writerow(["Rows", int(totals.get("rows") or 0)])
    writer.writerow(["Rows total", int(totals.get("rows_total") or totals.get("rows") or 0)])
    writer.writerow(["Rows returned", int(totals.get("rows_returned") or totals.get("rows") or 0)])
    writer.writerow(["Cantidad total", f"{Decimal(str(totals.get('cantidad_total') or 0)):.3f}"])
    writer.writerow(["MES", int(by_alcance.get(SolicitudVenta.ALCANCE_MES) or 0)])
    writer.writerow(["SEMANA", int(by_alcance.get(SolicitudVenta.ALCANCE_SEMANA) or 0)])
    writer.writerow(["FIN_SEMANA", int(by_alcance.get(SolicitudVenta.ALCANCE_FIN_SEMANA) or 0)])
    writer.writerow(["Forecast status SOBRE", int(forecast_counts.get("SOBRE") or 0)])
    writer.writerow(["Forecast status BAJO", int(forecast_counts.get("BAJO") or 0)])
    writer.writerow(["Forecast status OK", int(forecast_counts.get("OK") or 0)])
    writer.writerow(["Forecast status SIN_FORECAST", int(forecast_counts.get("SIN_FORECAST") or 0)])
    writer.writerow([])
    writer.writerow(["DETALLE"])
    writer.writerow(
        [
            "id",
            "receta_id",
            "receta",
            "codigo_point",
            "sucursal_id",
            "sucursal",
            "sucursal_codigo",
            "alcance",
            "periodo",
            "fecha_inicio",
            "fecha_fin",
            "cantidad",
            "forecast_status",
            "forecast_qty",
            "delta_solicitud_vs_forecast",
            "fuente",
            "actualizado_en",
        ]
    )
    for row in items:
        forecast_ref = row.get("forecast_ref") or {}
        writer.writerow(
            [
                int(row.get("id") or 0),
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                row.get("codigo_point") or "",
                int(row.get("sucursal_id") or 0) if row.get("sucursal_id") else "",
                row.get("sucursal") or "",
                row.get("sucursal_codigo") or "",
                row.get("alcance") or "",
                row.get("periodo") or "",
                row.get("fecha_inicio") or "",
                row.get("fecha_fin") or "",
                f"{Decimal(str(row.get('cantidad') or 0)):.3f}",
                forecast_ref.get("status") or "",
                f"{Decimal(str(forecast_ref.get('forecast_qty') or 0)):.3f}",
                f"{Decimal(str(forecast_ref.get('delta_solicitud_vs_forecast') or 0)):.3f}",
                row.get("fuente") or "",
                str(row.get("actualizado_en") or ""),
            ]
        )
    return response


def _ventas_solicitud_aplicar_forecast_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    scope = payload.get("scope") or {}
    updated = payload.get("updated") or {}
    adjusted_rows = payload.get("adjusted_rows") or []
    compare = payload.get("compare_solicitud") or {}
    compare_totals = compare.get("totals") or {}
    compare_rows = compare.get("rows") or []
    periodo = str(scope.get("periodo") or _normalize_periodo_mes(None)).replace("-", "")
    sucursal_id = int(scope.get("sucursal_id") or 0)
    filename = f"ventas_solicitud_aplicar_forecast_{periodo}_{sucursal_id}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Alcance", scope.get("alcance") or ""])
        ws_resumen.append(["Periodo", scope.get("periodo") or ""])
        ws_resumen.append(["Inicio", scope.get("target_start") or ""])
        ws_resumen.append(["Fin", scope.get("target_end") or ""])
        ws_resumen.append(["Sucursal ID", sucursal_id if sucursal_id > 0 else None])
        ws_resumen.append(["Sucursal", scope.get("sucursal_nombre") or ""])
        ws_resumen.append(["Modo", scope.get("modo") or ""])
        ws_resumen.append(["Escenario", scope.get("escenario") or "base"])
        ws_resumen.append(["Confianza minima %", float(scope.get("min_confianza_pct") or 0)])
        ws_resumen.append(["Dry run", "SI" if updated.get("dry_run") else "NO"])
        ws_resumen.append(["Created", int(updated.get("created") or 0)])
        ws_resumen.append(["Updated", int(updated.get("updated") or 0)])
        ws_resumen.append(["Skipped", int(updated.get("skipped") or 0)])
        ws_resumen.append(["Skipped cap", int(updated.get("skipped_cap") or 0)])
        ws_resumen.append(["Applied", int(updated.get("applied") or 0)])
        ws_resumen.append(["Compare forecast total", float(compare_totals.get("forecast_total") or 0)])
        ws_resumen.append(["Compare solicitud total", float(compare_totals.get("solicitud_total") or 0)])
        ws_resumen.append(["Compare delta total", float(compare_totals.get("delta_total") or 0)])
        ws_resumen.append(["Compare en rango", int(compare_totals.get("en_rango_count") or 0)])
        ws_resumen.append(["Compare sobre rango", int(compare_totals.get("sobre_rango_count") or 0)])
        ws_resumen.append(["Compare bajo rango", int(compare_totals.get("bajo_rango_count") or 0)])

        ws_ajustes = wb.create_sheet("Ajustes")
        ws_ajustes.append(
            [
                "Receta ID",
                "Receta",
                "Cantidad anterior",
                "Cantidad nueva",
                "Variacion %",
                "Accion",
                "Status previo",
            ]
        )
        for row in adjusted_rows:
            ws_ajustes.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    float(row.get("anterior") or 0),
                    float(row.get("nueva") or 0),
                    float(row.get("variacion_pct")) if row.get("variacion_pct") is not None else None,
                    row.get("accion") or "",
                    row.get("status_before") or "",
                ]
            )

        ws_compare = wb.create_sheet("Compare")
        ws_compare.append(
            [
                "Receta ID",
                "Receta",
                "Forecast",
                "Forecast base",
                "Forecast bajo",
                "Forecast alto",
                "Solicitud",
                "Delta",
                "Variacion %",
                "Status",
                "Status rango",
            ]
        )
        for row in compare_rows:
            ws_compare.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    float(row.get("forecast_qty") or 0),
                    float(row.get("forecast_base") or 0),
                    float(row.get("forecast_low") or 0),
                    float(row.get("forecast_high") or 0),
                    float(row.get("solicitud_qty") or 0),
                    float(row.get("delta_qty") or 0),
                    float(row.get("variacion_pct")) if row.get("variacion_pct") is not None else None,
                    row.get("status") or "",
                    row.get("status_rango") or "",
                ]
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Alcance", scope.get("alcance") or ""])
    writer.writerow(["Periodo", scope.get("periodo") or ""])
    writer.writerow(["Inicio", scope.get("target_start") or ""])
    writer.writerow(["Fin", scope.get("target_end") or ""])
    writer.writerow(["Sucursal ID", sucursal_id if sucursal_id > 0 else ""])
    writer.writerow(["Sucursal", scope.get("sucursal_nombre") or ""])
    writer.writerow(["Modo", scope.get("modo") or ""])
    writer.writerow(["Escenario", scope.get("escenario") or "base"])
    writer.writerow(["Confianza minima %", f"{Decimal(str(scope.get('min_confianza_pct') or 0)):.1f}"])
    writer.writerow(["Dry run", "SI" if updated.get("dry_run") else "NO"])
    writer.writerow(["Created", int(updated.get("created") or 0)])
    writer.writerow(["Updated", int(updated.get("updated") or 0)])
    writer.writerow(["Skipped", int(updated.get("skipped") or 0)])
    writer.writerow(["Skipped cap", int(updated.get("skipped_cap") or 0)])
    writer.writerow(["Applied", int(updated.get("applied") or 0)])
    writer.writerow(["Compare forecast total", f"{Decimal(str(compare_totals.get('forecast_total') or 0)):.3f}"])
    writer.writerow(["Compare solicitud total", f"{Decimal(str(compare_totals.get('solicitud_total') or 0)):.3f}"])
    writer.writerow(["Compare delta total", f"{Decimal(str(compare_totals.get('delta_total') or 0)):.3f}"])
    writer.writerow(["Compare en rango", int(compare_totals.get("en_rango_count") or 0)])
    writer.writerow(["Compare sobre rango", int(compare_totals.get("sobre_rango_count") or 0)])
    writer.writerow(["Compare bajo rango", int(compare_totals.get("bajo_rango_count") or 0)])
    writer.writerow([])
    writer.writerow(["AJUSTES"])
    writer.writerow(["receta_id", "receta", "anterior", "nueva", "variacion_pct", "accion", "status_before"])
    for row in adjusted_rows:
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                f"{Decimal(str(row.get('anterior') or 0)):.3f}",
                f"{Decimal(str(row.get('nueva') or 0)):.3f}",
                f"{Decimal(str(row.get('variacion_pct'))):.1f}" if row.get("variacion_pct") is not None else "",
                row.get("accion") or "",
                row.get("status_before") or "",
            ]
        )
    writer.writerow([])
    writer.writerow(["COMPARE_SOLICITUD"])
    writer.writerow(
        [
            "receta_id",
            "receta",
            "forecast",
            "forecast_base",
            "forecast_bajo",
            "forecast_alto",
            "solicitud",
            "delta",
            "variacion_pct",
            "status",
            "status_rango",
        ]
    )
    for row in compare_rows:
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                f"{Decimal(str(row.get('forecast_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_base') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_low') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_high') or 0)):.3f}",
                f"{Decimal(str(row.get('solicitud_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('delta_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('variacion_pct'))):.1f}" if row.get("variacion_pct") is not None else "",
                row.get("status") or "",
                row.get("status_rango") or "",
            ]
        )
    return response


def _resolve_receta_bulk_ref(
    *,
    receta_id: int | None,
    receta_name: str,
    codigo_point: str,
    cache: dict[tuple[int, str, str], Receta | None],
) -> Receta | None:
    receta_id_int = int(receta_id) if receta_id else 0
    code_norm = (codigo_point or "").strip().lower()
    name_norm = (receta_name or "").strip().lower()
    key = (receta_id_int, code_norm, name_norm)
    if key in cache:
        return cache[key]

    receta = None
    if receta_id_int > 0:
        receta = Receta.objects.filter(pk=receta_id_int).first()
    if receta is None:
        receta = _resolve_receta_for_sales(receta_name, codigo_point)
    cache[key] = receta
    return receta


def _resolve_sucursal_bulk_ref(
    *,
    sucursal_id: int | None,
    sucursal_name: str,
    sucursal_codigo: str,
    default_sucursal: Sucursal | None,
    cache: dict[tuple[int, str, str, int], Sucursal | None],
) -> Sucursal | None:
    sucursal_id_int = int(sucursal_id) if sucursal_id else 0
    code_norm = (sucursal_codigo or "").strip().lower()
    name_norm = (sucursal_name or "").strip().lower()
    default_id = int(default_sucursal.id) if default_sucursal else 0
    key = (sucursal_id_int, code_norm, name_norm, default_id)
    if key in cache:
        return cache[key]

    sucursal = None
    if sucursal_id_int > 0:
        sucursal = Sucursal.objects.filter(pk=sucursal_id_int, activa=True).first()
    else:
        sucursal = _resolve_sucursal_for_sales(sucursal_name, sucursal_codigo, default_sucursal)
        if sucursal and not sucursal.activa:
            sucursal = None
    cache[key] = sucursal
    return sucursal


def _month_start(value: date) -> date:
    return date(value.year, value.month, 1)


def _shift_month(start: date, months: int) -> date:
    month_index = (start.year * 12 + (start.month - 1)) + months
    year = month_index // 12
    month = (month_index % 12) + 1
    return date(year, month, 1)


def _forecast_backtest_windows(alcance: str, fecha_base: date, periods: int) -> list[tuple[date, date]]:
    windows: list[tuple[date, date]] = []
    if alcance == "mes":
        anchor = _month_start(fecha_base)
        for lag in range(periods, 0, -1):
            start = _shift_month(anchor, -lag)
            end = _shift_month(start, 1) - timedelta(days=1)
            windows.append((start, end))
        return windows

    if alcance == "fin_semana":
        wd = fecha_base.weekday()
        if wd == 5:
            anchor = fecha_base
        elif wd == 6:
            anchor = fecha_base - timedelta(days=1)
        else:
            anchor = fecha_base + timedelta(days=(5 - wd))
        for lag in range(periods, 0, -1):
            start = anchor - timedelta(days=7 * lag)
            windows.append((start, start + timedelta(days=1)))
        return windows

    # semana
    anchor = fecha_base - timedelta(days=fecha_base.weekday())
    for lag in range(periods, 0, -1):
        start = anchor - timedelta(days=7 * lag)
        windows.append((start, start + timedelta(days=6)))
    return windows


class MRPExplodeView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ser = MRPRequestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        receta_id = ser.validated_data["receta_id"]
        multiplicador: Decimal = ser.validated_data.get("multiplicador", Decimal("1"))

        receta = Receta.objects.get(pk=receta_id)
        agregados = {}

        for l in receta.lineas.select_related("insumo").all():
            key = l.insumo.nombre if l.insumo else f"(NO MATCH) {l.insumo_texto}"
            if key not in agregados:
                agregados[key] = {"insumo_id": l.insumo.id if l.insumo else None, "nombre": key, "cantidad": Decimal("0"), "unidad": l.unidad_texto, "costo": 0.0}
            qty = Decimal(str(l.cantidad or 0)) * multiplicador
            agregados[key]["cantidad"] += qty
            agregados[key]["costo"] += float(l.costo_total_estimado) * float(multiplicador)

        items = sorted(agregados.values(), key=lambda x: x["nombre"])
        costo_total = sum(i["costo"] for i in items)

        return Response({
            "receta_id": receta.id,
            "receta_nombre": receta.nombre,
            "multiplicador": str(multiplicador),
            "costo_total": costo_total,
            "items": [
                {
                    "insumo_id": i["insumo_id"],
                    "nombre": i["nombre"],
                    "cantidad": str(i["cantidad"]),
                    "unidad": i["unidad"],
                    "costo": i["costo"],
                } for i in items
            ],
        }, status=status.HTTP_200_OK)


class RecetaVersionesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, receta_id: int):
        receta = get_object_or_404(Receta, pk=receta_id)
        data_unavailable = False
        warnings: list[str] = []
        try:
            asegurar_version_costeo(receta, fuente="API_VERSIONES")
        except (OperationalError, ProgrammingError):
            data_unavailable = True
            warnings.append("Versionado automático no disponible en este entorno.")
        limit = _parse_bounded_int(
            request.GET.get("limit", 25),
            default=25,
            min_value=1,
            max_value=200,
        )
        try:
            versiones = _load_versiones_costeo(receta, limit)
        except (OperationalError, ProgrammingError):
            data_unavailable = True
            versiones = []
            warnings.append("Histórico de versiones no disponible en este entorno.")
        payload = RecetaCostoVersionSerializer(versiones, many=True).data
        return Response(
            {
                "receta_id": receta.id,
                "receta_nombre": receta.nombre,
                "total": len(payload),
                "items": payload,
                "data_unavailable": data_unavailable,
                "warnings": warnings,
            },
            status=status.HTTP_200_OK,
        )


class RecetaCostoHistoricoView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, receta_id: int):
        receta = get_object_or_404(Receta, pk=receta_id)
        data_unavailable = False
        warnings: list[str] = []
        try:
            asegurar_version_costeo(receta, fuente="API_HISTORICO")
        except (OperationalError, ProgrammingError):
            data_unavailable = True
            warnings.append("Versionado automático no disponible en este entorno.")
        limit = _parse_bounded_int(
            request.GET.get("limit", 60),
            default=60,
            min_value=1,
            max_value=300,
        )
        try:
            versiones = _load_versiones_costeo(receta, limit)
        except (OperationalError, ProgrammingError):
            data_unavailable = True
            versiones = []
            warnings.append("Histórico de costos no disponible en este entorno.")
        payload = RecetaCostoVersionSerializer(versiones, many=True).data
        comparativo = comparativo_versiones(versiones)
        data = {
            "receta_id": receta.id,
            "receta_nombre": receta.nombre,
            "puntos": payload,
            "data_unavailable": data_unavailable,
            "warnings": warnings,
        }
        if comparativo:
            data["comparativo"] = {
                "version_actual": comparativo["version_actual"],
                "version_previa": comparativo["version_previa"],
                "delta_total": str(comparativo["delta_total"]),
                "delta_pct": str(comparativo["delta_pct"]),
            }

        base_raw = (request.GET.get("base") or "").strip()
        target_raw = (request.GET.get("target") or "").strip()
        if base_raw and target_raw:
            try:
                base_num = int(base_raw)
                target_num = int(target_raw)
            except ValueError:
                base_num = None
                target_num = None

            if base_num is not None and target_num is not None and base_num != target_num:
                by_version = {v.version_num: v for v in versiones}
                base_v = by_version.get(base_num)
                target_v = by_version.get(target_num)
                if base_v and target_v:
                    base_total = Decimal(str(base_v.costo_total or 0))
                    target_total = Decimal(str(target_v.costo_total or 0))
                    delta_total = target_total - base_total
                    delta_pct = None
                    if base_total > 0:
                        delta_pct = (delta_total * Decimal("100")) / base_total

                    data["comparativo_seleccionado"] = {
                        "base": base_v.version_num,
                        "target": target_v.version_num,
                        "delta_mp": str(Decimal(str(target_v.costo_mp or 0)) - Decimal(str(base_v.costo_mp or 0))),
                        "delta_mo": str(Decimal(str(target_v.costo_mo or 0)) - Decimal(str(base_v.costo_mo or 0))),
                        "delta_indirecto": str(Decimal(str(target_v.costo_indirecto or 0)) - Decimal(str(base_v.costo_indirecto or 0))),
                        "delta_total": str(delta_total),
                        "delta_pct_total": str(delta_pct) if delta_pct is not None else None,
                    }
        return Response(data, status=status.HTTP_200_OK)


class InventarioSugerenciasCompraView(APIView):
    permission_classes = [IsAuthenticated]

    def _resolve_scope(self, request):
        plan_id_raw = (request.GET.get("plan_id") or "").strip()
        periodo_raw = (request.GET.get("periodo") or "").strip()
        plan = None
        year = None
        month = None

        if plan_id_raw:
            try:
                plan_id = int(plan_id_raw)
            except ValueError:
                plan_id = None
            if plan_id:
                plan = get_object_or_404(PlanProduccion, pk=plan_id)
                year = plan.fecha_produccion.year
                month = plan.fecha_produccion.month

        if not plan:
            parsed = _parse_period(periodo_raw)
            if parsed:
                year, month = parsed
            else:
                today = timezone.localdate()
                year, month = today.year, today.month

        if plan:
            items_qs = PlanProduccionItem.objects.filter(plan_id=plan.id).select_related("receta", "plan")
        else:
            items_qs = PlanProduccionItem.objects.filter(
                plan__fecha_produccion__year=year,
                plan__fecha_produccion__month=month,
            ).select_related("receta", "plan")

        return items_qs, {
            "plan_id": plan.id if plan else None,
            "plan_nombre": plan.nombre if plan else "",
            "periodo": f"{year:04d}-{month:02d}",
        }

    def _requerimientos_plan(self, plan_items_qs) -> dict[int, Decimal]:
        requerimientos: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
        receta_ids = set(plan_items_qs.values_list("receta_id", flat=True))
        if not receta_ids:
            return requerimientos

        lineas_by_receta: dict[int, list[LineaReceta]] = defaultdict(list)
        lineas_qs = (
            LineaReceta.objects.filter(receta_id__in=receta_ids)
            .exclude(tipo_linea=LineaReceta.TIPO_SUBSECCION)
            .only("receta_id", "insumo_id", "cantidad")
        )
        for linea in lineas_qs:
            if not linea.insumo_id:
                continue
            qty = _to_decimal(linea.cantidad)
            if qty <= 0:
                continue
            lineas_by_receta[linea.receta_id].append(linea)

        for item in plan_items_qs:
            factor = _to_decimal(item.cantidad)
            if factor <= 0:
                continue
            for linea in lineas_by_receta.get(item.receta_id, []):
                requerimientos[linea.insumo_id] += _to_decimal(linea.cantidad) * factor
        return requerimientos

    def _en_transito_por_insumo(self) -> dict[int, Decimal]:
        totals: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
        transit_status = [
            OrdenCompra.STATUS_ENVIADA,
            OrdenCompra.STATUS_CONFIRMADA,
            OrdenCompra.STATUS_PARCIAL,
        ]
        ordenes = (
            OrdenCompra.objects.filter(estatus__in=transit_status, solicitud__isnull=False)
            .select_related("solicitud__insumo")
            .only("solicitud__insumo_id", "solicitud__cantidad")
        )
        for orden in ordenes:
            solicitud = orden.solicitud
            if not solicitud or not solicitud.insumo_id:
                continue
            qty = _to_decimal(solicitud.cantidad)
            if qty > 0:
                totals[solicitud.insumo_id] += qty
        return totals

    def get(self, request):
        include_all = _parse_bool(request.GET.get("include_all"), default=False)
        try:
            limit = int(request.GET.get("limit", 300))
        except ValueError:
            limit = 300
        limit = max(1, min(limit, 2000))

        plan_items_qs, scope = self._resolve_scope(request)
        requerimientos = self._requerimientos_plan(plan_items_qs)
        en_transito = self._en_transito_por_insumo()

        existencias_qs = ExistenciaInsumo.objects.select_related("insumo__unidad_base", "insumo__proveedor_principal").filter(
            insumo__activo=True
        )
        existencias_map = {e.insumo_id: e for e in existencias_qs}

        insumo_ids = set(existencias_map.keys()) | set(requerimientos.keys()) | set(en_transito.keys())
        if not insumo_ids:
            return Response(
                {
                    "scope": scope,
                    "formula": "compra_sugerida = (requerido + stock_seguridad) - (disponible + en_transito)",
                    "totales": {
                        "insumos": 0,
                        "criticos": 0,
                        "bajo_reorden": 0,
                        "compra_sugerida_total": "0",
                        "costo_estimado_total": "0",
                    },
                    "items": [],
                },
                status=status.HTTP_200_OK,
            )

        insumos = {
            i.id: i
            for i in Insumo.objects.filter(id__in=insumo_ids, activo=True).select_related("unidad_base", "proveedor_principal")
        }
        if not insumos:
            return Response(
                {
                    "scope": scope,
                    "formula": "compra_sugerida = (requerido + stock_seguridad) - (disponible + en_transito)",
                    "totales": {
                        "insumos": 0,
                        "criticos": 0,
                        "bajo_reorden": 0,
                        "compra_sugerida_total": "0",
                        "costo_estimado_total": "0",
                    },
                    "items": [],
                },
                status=status.HTTP_200_OK,
            )

        latest_cost: dict[int, Decimal] = {}
        for costo in (
            CostoInsumo.objects.filter(insumo_id__in=list(insumos.keys()))
            .only("insumo_id", "costo_unitario", "fecha", "id")
            .order_by("insumo_id", "-fecha", "-id")
        ):
            if costo.insumo_id not in latest_cost:
                latest_cost[costo.insumo_id] = _to_decimal(costo.costo_unitario)

        rows = []
        total_sugerido = Decimal("0")
        total_costo = Decimal("0")
        criticos = 0
        bajo_reorden = 0

        for insumo_id in sorted(insumos.keys(), key=lambda pk: insumos[pk].nombre.lower()):
            insumo = insumos[insumo_id]
            ex = existencias_map.get(insumo_id)

            stock_actual = _to_decimal(ex.stock_actual if ex else 0)
            stock_seguridad = _to_decimal(ex.stock_minimo if ex else 0)
            punto_reorden = _to_decimal(ex.punto_reorden if ex else 0)
            consumo_diario = _to_decimal(ex.consumo_diario_promedio if ex else 0)
            lead_time = int(ex.dias_llegada_pedido or 0) if ex else 0
            if lead_time <= 0 and insumo.proveedor_principal_id:
                lead_time = int(insumo.proveedor_principal.lead_time_dias or 0)
            lead_time = max(lead_time, 0)

            demanda_lead_time = consumo_diario * Decimal(str(lead_time))
            requerido_plan = _to_decimal(requerimientos.get(insumo_id, 0))
            requerido = requerido_plan if requerido_plan > demanda_lead_time else demanda_lead_time
            en_transito_qty = _to_decimal(en_transito.get(insumo_id, 0))

            sugerida = requerido + stock_seguridad - (stock_actual + en_transito_qty)
            if sugerida < 0:
                sugerida = Decimal("0")

            costo_unitario = _to_decimal(latest_cost.get(insumo_id, 0))
            costo_sugerido = sugerida * costo_unitario if sugerida > 0 and costo_unitario > 0 else Decimal("0")

            if stock_actual <= 0 and (requerido > 0 or punto_reorden > 0):
                estado = "CRITICO"
                criticos += 1
            elif punto_reorden > 0 and stock_actual < punto_reorden:
                estado = "BAJO_REORDEN"
                bajo_reorden += 1
            else:
                estado = "SUFICIENTE"

            if not include_all and sugerida <= 0:
                continue

            total_sugerido += sugerida
            total_costo += costo_sugerido

            rows.append(
                {
                    "insumo_id": insumo_id,
                    "insumo": insumo.nombre,
                    "unidad": insumo.unidad_base.codigo if insumo.unidad_base_id and insumo.unidad_base else "",
                    "proveedor_principal": insumo.proveedor_principal.nombre if insumo.proveedor_principal_id else "",
                    "stock_actual": str(stock_actual),
                    "stock_seguridad": str(stock_seguridad),
                    "punto_reorden": str(punto_reorden),
                    "en_transito": str(en_transito_qty),
                    "requerido_plan": str(requerido_plan),
                    "demanda_lead_time": str(demanda_lead_time),
                    "requerido_total": str(requerido),
                    "compra_sugerida": str(sugerida),
                    "lead_time_dias": lead_time,
                    "consumo_diario_promedio": str(consumo_diario),
                    "costo_unitario": str(costo_unitario),
                    "costo_compra_sugerida": str(costo_sugerido),
                    "estatus": estado,
                }
            )

        rows.sort(
            key=lambda x: (
                Decimal(x["compra_sugerida"]),
                Decimal(x["requerido_total"]),
                x["insumo"].lower(),
            ),
            reverse=True,
        )

        return Response(
            {
                "scope": {**scope, "include_all": include_all, "limit": limit},
                "formula": "compra_sugerida = (requerido + stock_seguridad) - (disponible + en_transito)",
                "totales": {
                    "insumos": len(rows[:limit]),
                    "criticos": criticos,
                    "bajo_reorden": bajo_reorden,
                    "compra_sugerida_total": str(total_sugerido),
                    "costo_estimado_total": str(total_costo),
                },
                "items": rows[:limit],
            },
            status=status.HTTP_200_OK,
        )


class InventarioAliasesListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _serialize_alias(alias: InsumoAlias) -> dict:
        return {
            "id": alias.id,
            "alias": alias.nombre,
            "normalizado": alias.nombre_normalizado,
            "insumo_id": alias.insumo_id,
            "insumo": alias.insumo.nombre if alias.insumo_id else "",
            "unidad": alias.insumo.unidad_base.codigo if alias.insumo_id and alias.insumo.unidad_base_id else "",
            "categoria": alias.insumo.categoria if alias.insumo_id else "",
        }

    def get(self, request):
        if not can_view_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar aliases de inventario."},
                status=status.HTTP_403_FORBIDDEN,
            )

        q = (request.GET.get("q") or "").strip()
        insumo_id_raw = (request.GET.get("insumo_id") or "").strip()
        limit = _parse_bounded_int(request.GET.get("limit", 250), default=250, min_value=1, max_value=1200)

        qs = InsumoAlias.objects.select_related("insumo__unidad_base").order_by("nombre")
        if q:
            q_norm = normalizar_nombre(q)
            qs = qs.filter(
                Q(nombre__icontains=q)
                | Q(nombre_normalizado__icontains=q_norm)
                | Q(insumo__nombre__icontains=q)
                | Q(insumo__codigo_point__icontains=q)
            )

        insumo_id = None
        if insumo_id_raw:
            try:
                insumo_id = int(insumo_id_raw)
            except ValueError:
                return Response(
                    {"detail": "insumo_id inválido."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            qs = qs.filter(insumo_id=insumo_id)

        total_rows = qs.count()
        rows = [self._serialize_alias(a) for a in qs[:limit]]
        return Response(
            {
                "filters": {"q": q, "insumo_id": insumo_id, "limit": limit},
                "totales": {"rows": total_rows, "returned": len(rows)},
                "items": rows,
            },
            status=status.HTTP_200_OK,
        )

    def post(self, request):
        if not can_manage_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para crear/editar aliases de inventario."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = InventarioAliasCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        alias_name = (data.get("alias") or "").strip()
        alias_norm = normalizar_nombre(alias_name)
        if not alias_norm:
            return Response(
                {"detail": "Alias inválido: nombre vacío después de normalizar."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        insumo = Insumo.objects.filter(pk=data["insumo_id"], activo=True).select_related("unidad_base").first()
        if insumo is None:
            return Response(
                {"detail": "insumo_id no encontrado o inactivo."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        alias, created = InsumoAlias.objects.get_or_create(
            nombre_normalizado=alias_norm,
            defaults={"nombre": alias_name[:250], "insumo": insumo},
        )

        updated = False
        if not created:
            changed = []
            if alias.insumo_id != insumo.id:
                alias.insumo = insumo
                changed.append("insumo")
            if alias.nombre != alias_name[:250]:
                alias.nombre = alias_name[:250]
                changed.append("nombre")
            if changed:
                alias.save(update_fields=changed)
                updated = True

        point_resolved = 0
        recetas_resolved = 0
        if bool(data.get("resolver_cross_source", True)):
            point_resolved, recetas_resolved = _resolve_cross_source_with_alias(alias_name, insumo)

        alias.refresh_from_db()
        return Response(
            {
                "created": created,
                "updated": updated,
                "alias": self._serialize_alias(alias),
                "resolved": {
                    "point_pending": point_resolved,
                    "recetas_pending": recetas_resolved,
                },
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )


class InventarioAliasesMassReassignView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not can_manage_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para reasignar aliases de inventario."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = InventarioAliasMassReassignSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        insumo = Insumo.objects.filter(pk=data["insumo_id"], activo=True).select_related("unidad_base").first()
        if insumo is None:
            return Response(
                {"detail": "insumo_id no encontrado o inactivo."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        alias_ids = list(dict.fromkeys(int(aid) for aid in data["alias_ids"]))
        aliases = list(InsumoAlias.objects.filter(id__in=alias_ids).select_related("insumo"))
        if not aliases:
            return Response(
                {"detail": "No se encontraron aliases con los IDs enviados."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        resolver_cross = bool(data.get("resolver_cross_source", True))
        updated = 0
        point_resolved = 0
        recetas_resolved = 0
        touched_ids: list[int] = []
        for alias in aliases:
            if alias.insumo_id == insumo.id:
                continue
            alias.insumo = insumo
            alias.save(update_fields=["insumo"])
            updated += 1
            touched_ids.append(alias.id)
            if resolver_cross:
                p_count, r_count = _resolve_cross_source_with_alias(alias.nombre, insumo)
                point_resolved += p_count
                recetas_resolved += r_count

        return Response(
            {
                "target_insumo": {"id": insumo.id, "nombre": insumo.nombre},
                "selected": len(alias_ids),
                "found": len(aliases),
                "updated": updated,
                "touched_alias_ids": touched_ids,
                "resolved": {
                    "point_pending": point_resolved,
                    "recetas_pending": recetas_resolved,
                },
            },
            status=status.HTTP_200_OK,
        )


class InventarioAliasesPendientesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not can_view_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar pendientes de homologación."},
                status=status.HTTP_403_FORBIDDEN,
            )

        limit = _parse_bounded_int(request.GET.get("limit", 120), default=120, min_value=1, max_value=400)
        runs_to_scan = _parse_bounded_int(request.GET.get("runs", 5), default=5, min_value=1, max_value=30)
        point_tipo = (request.GET.get("point_tipo") or PointPendingMatch.TIPO_INSUMO).strip().upper()
        valid_point_tipos = {
            PointPendingMatch.TIPO_INSUMO,
            PointPendingMatch.TIPO_PROVEEDOR,
            PointPendingMatch.TIPO_PRODUCTO,
            "TODOS",
            "ALL",
        }
        if point_tipo not in valid_point_tipos:
            return Response(
                {"detail": "point_tipo inválido. Usa INSUMO, PROVEEDOR, PRODUCTO o TODOS."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        almacen_rows: list[dict] = []
        sync_runs = list(AlmacenSyncRun.objects.only("id", "started_at", "pending_preview").order_by("-started_at")[:runs_to_scan])
        for run in sync_runs:
            for row in run.pending_preview or []:
                nombre_origen = str((row or {}).get("nombre_origen") or "").strip()
                if not nombre_origen:
                    continue
                almacen_rows.append(
                    {
                        "run_id": run.id,
                        "run_started_at": run.started_at,
                        "nombre_origen": nombre_origen,
                        "nombre_normalizado": str((row or {}).get("nombre_normalizado") or normalizar_nombre(nombre_origen)),
                        "sugerencia": str((row or {}).get("suggestion") or ""),
                        "score": float((row or {}).get("score") or 0),
                        "metodo": str((row or {}).get("method") or ""),
                        "fuente": str((row or {}).get("fuente") or "ALMACEN"),
                    }
                )
                if len(almacen_rows) >= limit:
                    break
            if len(almacen_rows) >= limit:
                break

        point_qs = PointPendingMatch.objects.order_by("-fuzzy_score", "point_nombre")
        if point_tipo not in {"TODOS", "ALL"}:
            point_qs = point_qs.filter(tipo=point_tipo)
        point_total = point_qs.count()
        point_totals_by_tipo = {
            row["tipo"]: row["count"]
            for row in (
                point_qs.values("tipo")
                .annotate(count=Count("id"))
                .order_by("tipo")
            )
        }
        point_rows = [
            {
                "id": p.id,
                "tipo": p.tipo,
                "point_codigo": p.point_codigo,
                "point_nombre": p.point_nombre,
                "sugerencia": p.fuzzy_sugerencia or "",
                "score": float(p.fuzzy_score or 0),
                "metodo": p.method or "",
                "actualizado_en": p.actualizado_en,
            }
            for p in point_qs[:limit]
        ]

        recetas_qs = (
            LineaReceta.objects.filter(insumo__isnull=True)
            .filter(match_status__in=[LineaReceta.STATUS_NEEDS_REVIEW, LineaReceta.STATUS_REJECTED])
            .select_related("receta")
            .order_by("-match_score", "receta__nombre", "posicion")
        )
        recetas_total = recetas_qs.count()
        recetas_rows = [
            {
                "id": linea.id,
                "receta_id": linea.receta_id,
                "receta": linea.receta.nombre,
                "insumo_texto": linea.insumo_texto or "",
                "nombre_normalizado": normalizar_nombre(linea.insumo_texto or ""),
                "score": float(linea.match_score or 0),
                "metodo": linea.match_method or "",
                "estatus": linea.match_status,
            }
            for linea in recetas_qs[:limit]
        ]

        return Response(
            {
                "filters": {"limit": limit, "runs": runs_to_scan, "point_tipo": point_tipo},
                "totales": {
                    "almacen": len(almacen_rows),
                    "point": point_total,
                    "point_by_tipo": point_totals_by_tipo,
                    "recetas": recetas_total,
                },
                "items": {
                    "almacen": almacen_rows,
                    "point": point_rows,
                    "recetas": recetas_rows,
                },
            },
            status=status.HTTP_200_OK,
        )


class InventarioAliasesPendientesUnificadosView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not can_view_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar pendientes de homologación."},
                status=status.HTTP_403_FORBIDDEN,
            )

        limit = _parse_bounded_int(request.GET.get("limit", 120), default=120, min_value=1, max_value=600)
        runs_to_scan = _parse_bounded_int(request.GET.get("runs", 5), default=5, min_value=1, max_value=30)
        q = (request.GET.get("q") or "").strip()
        q_norm = normalizar_nombre(q)
        only_suggested = _parse_bool(request.GET.get("only_suggested"), default=False)
        min_sources = _parse_bounded_int(request.GET.get("min_sources", 1), default=1, min_value=1, max_value=3)
        score_min = float(_to_decimal(request.GET.get("score_min"), Decimal("0")))
        score_min = max(0.0, min(100.0, score_min))

        pending_rows: list[dict] = []
        sync_runs = list(AlmacenSyncRun.objects.only("id", "started_at", "pending_preview").order_by("-started_at")[:runs_to_scan])
        for run in sync_runs:
            for row in run.pending_preview or []:
                nombre_origen = str((row or {}).get("nombre_origen") or "").strip()
                if not nombre_origen:
                    continue
                pending_rows.append(
                    {
                        "nombre_origen": nombre_origen,
                        "nombre_normalizado": str((row or {}).get("nombre_normalizado") or normalizar_nombre(nombre_origen)),
                        "sugerencia": str((row or {}).get("suggestion") or ""),
                        "score": float((row or {}).get("score") or 0),
                        "source": str((row or {}).get("fuente") or "ALMACEN"),
                    }
                )

        pending_grouped = _build_pending_grouped(pending_rows)
        unified_rows, point_unmatched_count, receta_pending_lines = _build_cross_unified_rows(pending_grouped)
        filtered_rows = _apply_cross_filters(
            unified_rows,
            cross_q_norm=q_norm,
            cross_only_suggested=only_suggested,
            cross_min_sources=min_sources,
            cross_score_min=score_min,
        )

        overlap_2_plus = sum(1 for row in unified_rows if int(row.get("sources_active") or 0) >= 2)
        items = filtered_rows[:limit]

        return Response(
            {
                "filters": {
                    "limit": limit,
                    "runs": runs_to_scan,
                    "q": q,
                    "min_sources": min_sources,
                    "score_min": round(score_min, 2),
                    "only_suggested": only_suggested,
                },
                "totales": {
                    "runs_scanned": len(sync_runs),
                    "almacen_rows_raw": len(pending_rows),
                    "almacen_grouped": len(pending_grouped),
                    "unified_total": len(unified_rows),
                    "filtered_total": len(filtered_rows),
                    "returned": len(items),
                    "overlap_2_plus": overlap_2_plus,
                    "point_unmatched": point_unmatched_count,
                    "recetas_pending_lines": receta_pending_lines,
                },
                "items": items,
            },
            status=status.HTTP_200_OK,
        )


class InventarioAliasesPendientesUnificadosResolveView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not has_any_role(request.user, ROLE_ADMIN, ROLE_DG, ROLE_COMPRAS):
            return Response(
                {"detail": "No tienes permisos para resolver pendientes unificados."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = InventarioCrossPendientesResolveSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        limit = int(data.get("limit") or 300)
        runs_to_scan = int(data.get("runs") or 5)
        q = str(data.get("q") or "").strip()
        q_norm = normalizar_nombre(q)
        min_sources = int(data.get("min_sources") or 2)
        score_min = float(data.get("score_min") or 0)
        score_min = max(0.0, min(100.0, score_min))
        only_suggested = bool(data.get("only_suggested", True))
        dry_run = bool(data.get("dry_run", True))
        nombres = [str(x or "").strip() for x in (data.get("nombres") or [])]
        selected_norms = {normalizar_nombre(x) for x in nombres if normalizar_nombre(x)}

        pending_rows: list[dict] = []
        sync_runs = list(AlmacenSyncRun.objects.only("id", "started_at", "pending_preview").order_by("-started_at")[:runs_to_scan])
        for run in sync_runs:
            for row in run.pending_preview or []:
                nombre_origen = str((row or {}).get("nombre_origen") or "").strip()
                if not nombre_origen:
                    continue
                pending_rows.append(
                    {
                        "nombre_origen": nombre_origen,
                        "nombre_normalizado": str((row or {}).get("nombre_normalizado") or normalizar_nombre(nombre_origen)),
                        "sugerencia": str((row or {}).get("suggestion") or ""),
                        "score": float((row or {}).get("score") or 0),
                        "source": str((row or {}).get("fuente") or "ALMACEN"),
                    }
                )

        pending_grouped = _build_pending_grouped(pending_rows)
        unified_rows, _, _ = _build_cross_unified_rows(pending_grouped)
        filtered_rows = _apply_cross_filters(
            unified_rows,
            cross_q_norm=q_norm,
            cross_only_suggested=only_suggested,
            cross_min_sources=min_sources,
            cross_score_min=score_min,
        )
        if selected_norms:
            filtered_rows = [row for row in filtered_rows if (row.get("nombre_normalizado") or "") in selected_norms]
        rows_to_process = filtered_rows[:limit]

        processed = 0
        resolved = 0
        created_aliases = 0
        updated_aliases = 0
        unchanged = 0
        skipped_no_suggestion = 0
        skipped_no_target = 0
        point_resolved_total = 0
        recetas_resolved_total = 0
        preview_actions: list[dict] = []

        write_context = nullcontext()
        if not dry_run:
            write_context = transaction.atomic()

        with write_context:
            for row in rows_to_process:
                processed += 1
                alias_name = str(row.get("nombre_muestra") or "").strip()
                alias_norm = normalizar_nombre(alias_name)
                suggestion_name = str(row.get("suggestion") or "").strip()
                suggestion_norm = normalizar_nombre(suggestion_name)
                if not suggestion_norm:
                    skipped_no_suggestion += 1
                    if len(preview_actions) < 120:
                        preview_actions.append(
                            {
                                "nombre_muestra": alias_name,
                                "sugerencia": suggestion_name,
                                "action": "skip_no_suggestion",
                            }
                        )
                    continue

                insumo_target = Insumo.objects.filter(activo=True, nombre_normalizado=suggestion_norm).first()
                if not insumo_target:
                    skipped_no_target += 1
                    if len(preview_actions) < 120:
                        preview_actions.append(
                            {
                                "nombre_muestra": alias_name,
                                "sugerencia": suggestion_name,
                                "action": "skip_no_target",
                            }
                        )
                    continue

                action_name = "noop"
                if not alias_norm or alias_norm == insumo_target.nombre_normalizado:
                    unchanged += 1
                    action_name = "noop_same_name"
                else:
                    alias_obj = InsumoAlias.objects.filter(nombre_normalizado=alias_norm).first()
                    if dry_run:
                        if alias_obj is None:
                            action_name = "create_alias"
                        elif alias_obj.insumo_id != insumo_target.id or alias_obj.nombre != alias_name[:250]:
                            action_name = "update_alias"
                        else:
                            action_name = "noop_alias_exists"
                            unchanged += 1
                    else:
                        if alias_obj is None:
                            InsumoAlias.objects.create(
                                nombre=alias_name[:250],
                                nombre_normalizado=alias_norm,
                                insumo=insumo_target,
                            )
                            created_aliases += 1
                            action_name = "create_alias"
                        else:
                            changes = []
                            if alias_obj.insumo_id != insumo_target.id:
                                alias_obj.insumo = insumo_target
                                changes.append("insumo")
                            if alias_obj.nombre != alias_name[:250]:
                                alias_obj.nombre = alias_name[:250]
                                changes.append("nombre")
                            if changes:
                                alias_obj.save(update_fields=changes)
                                updated_aliases += 1
                                action_name = "update_alias"
                            else:
                                unchanged += 1
                                action_name = "noop_alias_exists"

                if not dry_run:
                    p_count, r_count = _resolve_cross_source_with_alias(alias_name or suggestion_name, insumo_target)
                    point_resolved_total += p_count
                    recetas_resolved_total += r_count

                resolved += 1
                if len(preview_actions) < 120:
                    preview_actions.append(
                        {
                            "nombre_muestra": alias_name,
                            "sugerencia": suggestion_name,
                            "insumo_target": insumo_target.nombre,
                            "insumo_id": insumo_target.id,
                            "action": action_name,
                            "sources_active": int(row.get("sources_active") or 0),
                            "total_count": int(row.get("total_count") or 0),
                        }
                    )

        return Response(
            {
                "dry_run": dry_run,
                "filters": {
                    "q": q,
                    "runs": runs_to_scan,
                    "limit": limit,
                    "min_sources": min_sources,
                    "score_min": round(score_min, 2),
                    "only_suggested": only_suggested,
                    "selected_names_count": len(selected_norms),
                },
                "totales": {
                    "candidatos_filtrados": len(filtered_rows),
                    "procesados": processed,
                    "resueltos": resolved,
                    "aliases_creados": created_aliases,
                    "aliases_actualizados": updated_aliases,
                    "sin_cambio": unchanged,
                    "sin_sugerencia": skipped_no_suggestion,
                    "sin_insumo_objetivo": skipped_no_target,
                    "point_resueltos": point_resolved_total,
                    "recetas_resueltas": recetas_resolved_total,
                },
                "items": preview_actions,
            },
            status=status.HTTP_200_OK,
        )


class IntegracionesDeactivateIdleClientsView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not can_view_audit(request.user):
            return Response(
                {"detail": "No tienes permisos para ejecutar operaciones de integración."},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = IntegracionesDeactivateIdleClientsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        idle_days = serializer.validated_data["idle_days"]
        limit = serializer.validated_data["limit"]
        dry_run = bool(serializer.validated_data.get("dry_run"))
        summary = (
            _preview_deactivate_idle_api_clients(idle_days=idle_days, limit=limit)
            if dry_run
            else _deactivate_idle_api_clients(idle_days=idle_days, limit=limit)
        )
        if "dry_run" not in summary:
            summary["dry_run"] = dry_run
        log_event(
            request.user,
            "PREVIEW_DEACTIVATE_IDLE_API_CLIENTS" if dry_run else "DEACTIVATE_IDLE_API_CLIENTS",
            "integraciones.PublicApiClient",
            "",
            payload=summary,
        )
        return Response(
            {
                "action": "deactivate_idle_clients",
                "summary": summary,
            },
            status=status.HTTP_200_OK,
        )


class IntegracionesPurgeApiLogsView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not can_view_audit(request.user):
            return Response(
                {"detail": "No tienes permisos para ejecutar operaciones de integración."},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = IntegracionesPurgeApiLogsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        retain_days = serializer.validated_data["retain_days"]
        max_delete = serializer.validated_data["max_delete"]
        dry_run = bool(serializer.validated_data.get("dry_run"))
        summary = (
            _preview_purge_api_logs(retain_days=retain_days, max_delete=max_delete)
            if dry_run
            else _purge_api_logs(retain_days=retain_days, max_delete=max_delete)
        )
        if "dry_run" not in summary:
            summary["dry_run"] = dry_run
        log_event(
            request.user,
            "PREVIEW_PURGE_API_LOGS" if dry_run else "PURGE_API_LOGS",
            "integraciones.PublicApiAccessLog",
            "",
            payload=summary,
        )
        return Response(
            {
                "action": "purge_api_logs",
                "summary": summary,
            },
            status=status.HTTP_200_OK,
        )


class IntegracionesMaintenanceRunView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not can_view_audit(request.user):
            return Response(
                {"detail": "No tienes permisos para ejecutar operaciones de integración."},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = IntegracionesMaintenanceRunSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        dry_run = bool(data.get("dry_run"))

        deactivate_summary = (
            _preview_deactivate_idle_api_clients(idle_days=data["idle_days"], limit=data["idle_limit"])
            if dry_run
            else _deactivate_idle_api_clients(idle_days=data["idle_days"], limit=data["idle_limit"])
        )
        purge_summary = (
            _preview_purge_api_logs(retain_days=data["retain_days"], max_delete=data["max_delete"])
            if dry_run
            else _purge_api_logs(retain_days=data["retain_days"], max_delete=data["max_delete"])
        )
        payload = {
            "dry_run": dry_run,
            "deactivate_idle_clients": deactivate_summary,
            "purge_api_logs": purge_summary,
        }
        log_event(
            request.user,
            "PREVIEW_RUN_API_MAINTENANCE" if dry_run else "RUN_API_MAINTENANCE",
            "integraciones.Operaciones",
            "",
            payload=payload,
        )
        return Response(payload, status=status.HTTP_200_OK)


class IntegracionPointResumenView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not can_view_maestros(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar resumen de integración Point."},
                status=status.HTTP_403_FORBIDDEN,
            )

        insumos_activos_qs = Insumo.objects.filter(activo=True)
        insumos_activos = insumos_activos_qs.count()
        insumos_con_codigo = insumos_activos_qs.exclude(Q(codigo_point="") | Q(codigo_point__isnull=True)).count()
        insumos_sin_codigo = max(insumos_activos - insumos_con_codigo, 0)
        insumos_cobertura = round((insumos_con_codigo * 100.0 / insumos_activos), 2) if insumos_activos else 100.0

        recetas_total = Receta.objects.count()
        receta_ids_primary = set(
            Receta.objects.exclude(Q(codigo_point="") | Q(codigo_point__isnull=True)).values_list("id", flat=True)
        )
        receta_ids_alias = set(
            RecetaCodigoPointAlias.objects.filter(activo=True).values_list("receta_id", flat=True)
        )
        recetas_homologadas_ids = receta_ids_primary.union(receta_ids_alias)
        recetas_homologadas = len(recetas_homologadas_ids)
        recetas_sin_homologar = max(recetas_total - recetas_homologadas, 0)
        recetas_cobertura = round((recetas_homologadas * 100.0 / recetas_total), 2) if recetas_total else 100.0

        point_pending_by_tipo = {
            row["tipo"]: row["count"]
            for row in (
                PointPendingMatch.objects.values("tipo")
                .annotate(count=Count("id"))
                .order_by("tipo")
            )
        }
        point_pending_total = sum(point_pending_by_tipo.values())

        latest_run = AlmacenSyncRun.objects.only("id", "started_at", "pending_preview").order_by("-started_at").first()
        almacen_pending_count = len((latest_run.pending_preview or [])) if latest_run else 0

        recetas_pending_lines = (
            LineaReceta.objects.filter(insumo__isnull=True)
            .filter(match_status__in=[LineaReceta.STATUS_NEEDS_REVIEW, LineaReceta.STATUS_REJECTED])
            .count()
        )

        proveedores_activos = Proveedor.objects.filter(activo=True).count()
        proveedores_pending_point = int(point_pending_by_tipo.get(PointPendingMatch.TIPO_PROVEEDOR, 0))
        now_dt = timezone.now()
        since_24h = now_dt - timedelta(hours=24)
        since_48h = now_dt - timedelta(hours=48)
        current_window_qs = PublicApiAccessLog.objects.filter(created_at__gte=since_24h)
        previous_window_qs = PublicApiAccessLog.objects.filter(created_at__gte=since_48h, created_at__lt=since_24h)
        requests_24h = current_window_qs.count()
        errors_24h = current_window_qs.filter(status_code__gte=400).count()
        requests_prev_24h = previous_window_qs.count()
        errors_prev_24h = previous_window_qs.filter(status_code__gte=400).count()
        requests_delta_pct = _pct_change(requests_24h, requests_prev_24h)
        errors_delta_pct = _pct_change(errors_24h, errors_prev_24h)
        errors_by_endpoint_24h = list(
            PublicApiAccessLog.objects.filter(created_at__gte=since_24h, status_code__gte=400)
            .values("endpoint")
            .annotate(
                total=Count("id"),
                clientes=Count("client_id", distinct=True),
                last_at=Max("created_at"),
            )
            .order_by("-total", "endpoint")[:12]
        )
        errors_by_client_24h = list(
            PublicApiAccessLog.objects.filter(created_at__gte=since_24h, status_code__gte=400)
            .values("client__nombre")
            .annotate(
                total=Count("id"),
                last_at=Max("created_at"),
            )
            .order_by("-total", "client__nombre")[:12]
        )
        api_daily_trend = _build_public_api_daily_trend(days=7)
        api_7d_requests = sum(int(row.get("requests") or 0) for row in api_daily_trend)
        api_7d_errors = sum(int(row.get("errors") or 0) for row in api_daily_trend)
        api_7d_error_rate = round((api_7d_errors * 100.0 / api_7d_requests), 2) if api_7d_requests else 0.0
        since_30d = timezone.now() - timedelta(days=30)
        api_clients_total = PublicApiClient.objects.count()
        api_clients_active = PublicApiClient.objects.filter(activo=True).count()
        api_clients_inactive = max(api_clients_total - api_clients_active, 0)
        clients_with_activity_30d = set(
            PublicApiAccessLog.objects.filter(created_at__gte=since_30d)
            .values_list("client_id", flat=True)
            .distinct()
        )
        api_clients_unused_30d = PublicApiClient.objects.filter(activo=True).exclude(id__in=clients_with_activity_30d).count()
        api_clients_top_30d = list(
            PublicApiAccessLog.objects.filter(created_at__gte=since_30d)
            .values("client__nombre")
            .annotate(
                requests=Count("id"),
                errors=Count("id", filter=Q(status_code__gte=400)),
            )
            .order_by("-requests", "client__nombre")[:10]
        )

        stale_limit = timezone.now() - timedelta(hours=24)
        alertas_operativas = []
        if errors_24h:
            alertas_operativas.append(
                {
                    "nivel": "danger",
                    "titulo": "Errores API en últimas 24h",
                    "detalle": f"{errors_24h} requests con status >= 400.",
                }
            )
        if errors_24h >= 5 and errors_delta_pct >= 50:
            alertas_operativas.append(
                {
                    "nivel": "danger",
                    "titulo": "Spike de errores API (24h)",
                    "detalle": (
                        f"Errores 24h: {errors_24h} vs {errors_prev_24h} previos "
                        f"({errors_delta_pct:+.2f}%)."
                    ),
                }
            )
        if point_pending_total:
            alertas_operativas.append(
                {
                    "nivel": "warning",
                    "titulo": "Pendientes Point abiertos",
                    "detalle": (
                        f"Total {point_pending_total}. "
                        f"Insumos {point_pending_by_tipo.get(PointPendingMatch.TIPO_INSUMO, 0)}, "
                        f"productos {point_pending_by_tipo.get(PointPendingMatch.TIPO_PRODUCTO, 0)}, "
                        f"proveedores {point_pending_by_tipo.get(PointPendingMatch.TIPO_PROVEEDOR, 0)}."
                    ),
                }
            )
        if recetas_pending_lines:
            alertas_operativas.append(
                {
                    "nivel": "warning",
                    "titulo": "Líneas receta sin match",
                    "detalle": f"{recetas_pending_lines} líneas requieren homologación interna.",
                }
            )
        if not latest_run:
            alertas_operativas.append(
                {
                    "nivel": "warning",
                    "titulo": "Sync de almacén no ejecutado",
                    "detalle": "No hay corridas de sincronización registradas.",
                }
            )
        elif latest_run.started_at and latest_run.started_at < stale_limit:
            alertas_operativas.append(
                {
                    "nivel": "warning",
                    "titulo": "Sync de almacén desactualizado",
                    "detalle": f"Último sync: {latest_run.started_at:%Y-%m-%d %H:%M}.",
                }
            )
        if not alertas_operativas:
            alertas_operativas.append(
                {
                    "nivel": "ok",
                    "titulo": "Operación estable",
                    "detalle": "Sin alertas críticas en integración, match y sincronización.",
                }
            )

        return Response(
            {
                "generated_at": timezone.now(),
                "api_24h": {
                    "requests": requests_24h,
                    "errors": errors_24h,
                    "errors_by_endpoint": errors_by_endpoint_24h,
                    "errors_by_client": errors_by_client_24h,
                },
                "api_24h_comparativo": {
                    "requests_prev_24h": requests_prev_24h,
                    "errors_prev_24h": errors_prev_24h,
                    "requests_delta_pct": requests_delta_pct,
                    "errors_delta_pct": errors_delta_pct,
                },
                "api_7d": {
                    "requests": api_7d_requests,
                    "errors": api_7d_errors,
                    "error_rate_pct": api_7d_error_rate,
                    "daily": api_daily_trend,
                },
                "api_clients": {
                    "total": api_clients_total,
                    "active": api_clients_active,
                    "inactive": api_clients_inactive,
                    "unused_30d": api_clients_unused_30d,
                    "top_30d": api_clients_top_30d,
                },
                "alertas_operativas": alertas_operativas,
                "insumos": {
                    "activos": insumos_activos,
                    "con_codigo_point": insumos_con_codigo,
                    "sin_codigo_point": insumos_sin_codigo,
                    "cobertura_pct": insumos_cobertura,
                },
                "recetas": {
                    "total": recetas_total,
                    "homologadas": recetas_homologadas,
                    "sin_homologar": recetas_sin_homologar,
                    "con_codigo_point_primario": len(receta_ids_primary),
                    "con_alias_codigo_point": len(receta_ids_alias),
                    "cobertura_pct": recetas_cobertura,
                },
                "proveedores": {
                    "activos": proveedores_activos,
                    "point_pending": proveedores_pending_point,
                },
                "point_pending": {
                    "total": point_pending_total,
                    "por_tipo": point_pending_by_tipo,
                },
                "inventario": {
                    "almacen_pending_preview": almacen_pending_count,
                    "almacen_latest_run_id": latest_run.id if latest_run else None,
                    "almacen_latest_run_at": latest_run.started_at if latest_run else None,
                    "recetas_pending_match": recetas_pending_lines,
                },
            },
            status=status.HTTP_200_OK,
        )


class InventarioPointPendingResolveView(APIView):
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _resolve_pending_insumo_row(
        pending: PointPendingMatch,
        insumo_target: Insumo,
        create_aliases_enabled: bool,
    ) -> tuple[bool, bool, int]:
        point_code = (pending.point_codigo or "").strip()
        if point_code and insumo_target.codigo_point and insumo_target.codigo_point != point_code:
            return False, True, 0

        changed = []
        if point_code and insumo_target.codigo_point != point_code:
            insumo_target.codigo_point = point_code[:80]
            changed.append("codigo_point")
        if insumo_target.nombre_point != pending.point_nombre:
            insumo_target.nombre_point = (pending.point_nombre or "")[:250]
            changed.append("nombre_point")
        if changed:
            insumo_target.save(update_fields=changed)

        alias_created = 0
        if create_aliases_enabled:
            alias_norm = normalizar_nombre(pending.point_nombre or "")
            if alias_norm and alias_norm != insumo_target.nombre_normalizado:
                alias, was_created = InsumoAlias.objects.get_or_create(
                    nombre_normalizado=alias_norm,
                    defaults={"nombre": (pending.point_nombre or "")[:250], "insumo": insumo_target},
                )
                if not was_created and alias.insumo_id != insumo_target.id:
                    alias.insumo = insumo_target
                    alias.save(update_fields=["insumo"])
                if was_created:
                    alias_created = 1

        pending.delete()
        return True, False, alias_created

    def post(self, request):
        if not has_any_role(request.user, ROLE_ADMIN, ROLE_COMPRAS):
            return Response(
                {"detail": "No tienes permisos para resolver pendientes Point."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = InventarioPointPendingResolveSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        action = data["action"]
        tipo = data["tipo"]
        pending_ids = list(dict.fromkeys(data.get("pending_ids") or []))
        q = (data.get("q") or "").strip()
        score_min = float(data.get("score_min", 90.0) or 0)
        create_aliases = bool(data.get("create_aliases", True))

        selected_qs = PointPendingMatch.objects.filter(tipo=tipo)
        if pending_ids:
            selected_qs = selected_qs.filter(id__in=pending_ids)
        elif action == InventarioPointPendingResolveSerializer.ACTION_AUTO_RESOLVE_INSUMOS:
            if q:
                selected_qs = selected_qs.filter(
                    Q(point_nombre__icontains=q)
                    | Q(point_codigo__icontains=q)
                    | Q(fuzzy_sugerencia__icontains=q)
                )
        selected_qs = selected_qs.order_by("-fuzzy_score", "point_nombre")

        selected = list(selected_qs)
        if not selected:
            return Response(
                {"detail": "No se encontraron pendientes para procesar con los filtros enviados."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        resolved = 0
        conflicts = 0
        aliases_created = 0
        skipped_low_score = 0
        skipped_no_suggestion = 0
        skipped_no_target = 0
        proveedores_created = 0

        if action == InventarioPointPendingResolveSerializer.ACTION_RESOLVE_INSUMOS:
            insumo_target = Insumo.objects.filter(pk=data.get("insumo_id"), activo=True).first()
            if insumo_target is None:
                return Response(
                    {"detail": "insumo_id no encontrado o inactivo."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            for pending in selected:
                row_resolved, row_conflict, row_alias_created = self._resolve_pending_insumo_row(
                    pending=pending,
                    insumo_target=insumo_target,
                    create_aliases_enabled=create_aliases,
                )
                if row_conflict:
                    conflicts += 1
                    continue
                if row_resolved:
                    resolved += 1
                    aliases_created += row_alias_created

            return Response(
                {
                    "action": action,
                    "tipo": tipo,
                    "selected": len(selected),
                    "resolved": resolved,
                    "conflicts": conflicts,
                    "aliases_created": aliases_created,
                    "target": {"insumo_id": insumo_target.id, "insumo": insumo_target.nombre},
                },
                status=status.HTTP_200_OK,
            )

        if action == InventarioPointPendingResolveSerializer.ACTION_AUTO_RESOLVE_INSUMOS:
            if tipo != PointPendingMatch.TIPO_INSUMO:
                return Response(
                    {"detail": "La auto-resolución por sugerencia aplica solo para tipo=INSUMO."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            for pending in selected:
                if float(pending.fuzzy_score or 0.0) < score_min:
                    skipped_low_score += 1
                    continue

                sugerencia_norm = normalizar_nombre(pending.fuzzy_sugerencia or "")
                if not sugerencia_norm:
                    skipped_no_suggestion += 1
                    continue

                insumo_target = (
                    Insumo.objects.filter(activo=True, nombre_normalizado=sugerencia_norm)
                    .only("id", "codigo_point", "nombre_point", "nombre_normalizado")
                    .first()
                )
                if not insumo_target:
                    skipped_no_target += 1
                    continue

                row_resolved, row_conflict, row_alias_created = self._resolve_pending_insumo_row(
                    pending=pending,
                    insumo_target=insumo_target,
                    create_aliases_enabled=create_aliases,
                )
                if row_conflict:
                    conflicts += 1
                    continue
                if row_resolved:
                    resolved += 1
                    aliases_created += row_alias_created

            return Response(
                {
                    "action": action,
                    "tipo": tipo,
                    "selected": len(selected),
                    "resolved": resolved,
                    "conflicts": conflicts,
                    "aliases_created": aliases_created,
                    "skipped": {
                        "low_score": skipped_low_score,
                        "no_suggestion": skipped_no_suggestion,
                        "no_target": skipped_no_target,
                    },
                    "score_min": round(score_min, 2),
                },
                status=status.HTTP_200_OK,
            )

        if action == InventarioPointPendingResolveSerializer.ACTION_RESOLVE_PRODUCTOS:
            receta_target = Receta.objects.filter(pk=data.get("receta_id")).first()
            if receta_target is None:
                return Response(
                    {"detail": "receta_id no encontrada."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            for pending in selected:
                point_code = (pending.point_codigo or "").strip()
                if point_code:
                    point_norm = normalizar_codigo_point(point_code)
                    primary_norm = normalizar_codigo_point(receta_target.codigo_point)
                    if not receta_target.codigo_point:
                        receta_target.codigo_point = point_code[:80]
                        receta_target.save(update_fields=["codigo_point"])
                    elif primary_norm != point_norm:
                        if not point_norm:
                            conflicts += 1
                            continue
                        if not create_aliases:
                            conflicts += 1
                            continue

                        alias, was_created = RecetaCodigoPointAlias.objects.get_or_create(
                            codigo_point_normalizado=point_norm,
                            defaults={
                                "receta": receta_target,
                                "codigo_point": point_code[:80],
                                "nombre_point": (pending.point_nombre or "")[:250],
                                "activo": True,
                            },
                        )
                        if not was_created and alias.receta_id != receta_target.id:
                            conflicts += 1
                            continue

                        changed = []
                        if alias.codigo_point != point_code[:80]:
                            alias.codigo_point = point_code[:80]
                            changed.append("codigo_point")
                        if (pending.point_nombre or "").strip() and alias.nombre_point != (pending.point_nombre or "")[:250]:
                            alias.nombre_point = (pending.point_nombre or "")[:250]
                            changed.append("nombre_point")
                        if not alias.activo:
                            alias.activo = True
                            changed.append("activo")
                        if changed:
                            alias.save(update_fields=changed)
                        if was_created:
                            aliases_created += 1

                pending.delete()
                resolved += 1

            return Response(
                {
                    "action": action,
                    "tipo": tipo,
                    "selected": len(selected),
                    "resolved": resolved,
                    "conflicts": conflicts,
                    "aliases_created": aliases_created,
                    "target": {"receta_id": receta_target.id, "receta": receta_target.nombre},
                },
                status=status.HTTP_200_OK,
            )

        if action == InventarioPointPendingResolveSerializer.ACTION_RESOLVE_PROVEEDORES:
            proveedor_target = None
            proveedor_id = data.get("proveedor_id")
            if proveedor_id:
                proveedor_target = Proveedor.objects.filter(pk=proveedor_id).first()
                if proveedor_target is None:
                    return Response(
                        {"detail": "proveedor_id no encontrado."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            for pending in selected:
                if proveedor_target is None:
                    _, was_created = Proveedor.objects.get_or_create(
                        nombre=(pending.point_nombre or "")[:200],
                        defaults={"activo": True},
                    )
                    if was_created:
                        proveedores_created += 1
                pending.delete()
                resolved += 1

            return Response(
                {
                    "action": action,
                    "tipo": tipo,
                    "selected": len(selected),
                    "resolved": resolved,
                    "proveedores_created": proveedores_created,
                    "target": (
                        {"proveedor_id": proveedor_target.id, "proveedor": proveedor_target.nombre}
                        if proveedor_target
                        else None
                    ),
                },
                status=status.HTTP_200_OK,
            )

        if action == InventarioPointPendingResolveSerializer.ACTION_DISCARD:
            deleted, _ = PointPendingMatch.objects.filter(id__in=[p.id for p in selected], tipo=tipo).delete()
            return Response(
                {
                    "action": action,
                    "tipo": tipo,
                    "selected": len(selected),
                    "discarded": deleted,
                },
                status=status.HTTP_200_OK,
            )

        return Response({"detail": "Acción no soportada."}, status=status.HTTP_400_BAD_REQUEST)


class InventarioAjustesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not can_view_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar ajustes de inventario."},
                status=status.HTTP_403_FORBIDDEN,
            )

        estatus = (request.GET.get("estatus") or "").strip().upper()
        valid_status = {
            AjusteInventario.STATUS_PENDIENTE,
            AjusteInventario.STATUS_APLICADO,
            AjusteInventario.STATUS_RECHAZADO,
        }
        qs = AjusteInventario.objects.select_related("insumo", "solicitado_por", "aprobado_por").order_by("-creado_en")
        if estatus in valid_status:
            qs = qs.filter(estatus=estatus)

        limit = _parse_bounded_int(
            request.GET.get("limit", 120),
            default=120,
            min_value=1,
            max_value=500,
        )
        items = [_serialize_ajuste_row(a) for a in qs[:limit]]

        totals_qs = qs if estatus in valid_status else AjusteInventario.objects.all()
        return Response(
            {
                "filters": {"estatus": estatus if estatus in valid_status else "", "limit": limit},
                "totales": {
                    "rows": len(items),
                    "pendientes": totals_qs.filter(estatus=AjusteInventario.STATUS_PENDIENTE).count(),
                    "aplicados": totals_qs.filter(estatus=AjusteInventario.STATUS_APLICADO).count(),
                    "rechazados": totals_qs.filter(estatus=AjusteInventario.STATUS_RECHAZADO).count(),
                },
                "items": items,
            },
            status=status.HTTP_200_OK,
        )

    def post(self, request):
        if not can_manage_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para registrar ajustes de inventario."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = InventarioAjusteCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        insumo = get_object_or_404(Insumo, pk=data["insumo_id"])
        aplicar_inmediato = bool(data.get("aplicar_inmediato"))
        comentario = data.get("comentario_revision") or ""
        if aplicar_inmediato and not _can_approve_ajustes(request.user):
            return Response(
                {
                    "detail": (
                        "No tienes permisos para aplicar ajustes inmediatamente. "
                        "Registra el ajuste en pendiente y solicita aprobación."
                    )
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        with transaction.atomic():
            ajuste = AjusteInventario.objects.create(
                insumo=insumo,
                cantidad_sistema=data["cantidad_sistema"],
                cantidad_fisica=data["cantidad_fisica"],
                motivo=data["motivo"],
                estatus=AjusteInventario.STATUS_PENDIENTE,
                solicitado_por=request.user,
            )
            if aplicar_inmediato:
                _apply_ajuste(ajuste, request.user, comentario=comentario)

        payload = _serialize_ajuste_row(ajuste)
        payload["aplicado"] = ajuste.estatus == AjusteInventario.STATUS_APLICADO
        return Response(payload, status=status.HTTP_201_CREATED)


class InventarioAjusteDecisionView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, ajuste_id: int):
        if not _can_approve_ajustes(request.user):
            return Response(
                {"detail": "No tienes permisos para aprobar/rechazar ajustes de inventario."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ajuste = get_object_or_404(
            AjusteInventario.objects.select_related("insumo", "solicitado_por", "aprobado_por"),
            pk=ajuste_id,
        )
        ser = InventarioAjusteDecisionSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        action = ser.validated_data["action"]
        comentario = ser.validated_data.get("comentario_revision") or ""

        if action == "reject":
            if ajuste.estatus == AjusteInventario.STATUS_APLICADO:
                return Response(
                    {"detail": "No se puede rechazar un ajuste ya aplicado."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            ajuste.estatus = AjusteInventario.STATUS_RECHAZADO
            ajuste.aprobado_por = request.user
            ajuste.aprobado_en = timezone.now()
            ajuste.aplicado_en = None
            ajuste.comentario_revision = comentario
            ajuste.save(
                update_fields=[
                    "estatus",
                    "aprobado_por",
                    "aprobado_en",
                    "aplicado_en",
                    "comentario_revision",
                ]
            )
            log_event(
                request.user,
                "REJECT",
                "inventario.AjusteInventario",
                ajuste.id,
                {"folio": ajuste.folio, "estatus": ajuste.estatus, "comentario_revision": comentario},
            )
        else:
            if ajuste.estatus == AjusteInventario.STATUS_APLICADO:
                return Response(
                    {
                        "detail": "El ajuste ya estaba aplicado.",
                        "item": _serialize_ajuste_row(ajuste),
                    },
                    status=status.HTTP_200_OK,
                )
            if ajuste.estatus == AjusteInventario.STATUS_RECHAZADO:
                return Response(
                    {"detail": "El ajuste fue rechazado y no puede aplicarse."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            _apply_ajuste(ajuste, request.user, comentario=comentario)

        payload = _serialize_ajuste_row(ajuste)
        payload["action"] = action
        return Response(payload, status=status.HTTP_200_OK)


class PresupuestosConsolidadoView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, periodo: str):
        if not can_view_compras(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar presupuestos de compras."},
                status=status.HTTP_403_FORBIDDEN,
            )

        parsed = _parse_period(periodo)
        if not parsed:
            return Response(
                {"detail": "Periodo inválido. Usa formato YYYY-MM."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        year, month = parsed
        periodo_mes = f"{year:04d}-{month:02d}"
        periodo_tipo = (request.GET.get("periodo_tipo") or "mes").strip().lower()
        if periodo_tipo not in {"mes", "q1", "q2"}:
            periodo_tipo = "mes"

        source_raw = request.GET.get("source")
        plan_raw = request.GET.get("plan_id")
        categoria_raw = request.GET.get("categoria")
        reabasto_raw = request.GET.get("reabasto")

        (
            solicitudes,
            source_filter,
            plan_filter,
            categoria_filter,
            reabasto_filter,
            _,
            _periodo_tipo,
            _periodo_mes,
            periodo_label,
        ) = _filtered_solicitudes(
            source_raw,
            plan_raw,
            categoria_raw,
            reabasto_raw,
            periodo_tipo,
            periodo_mes,
        )
        consumo_ref_filter = _sanitize_consumo_ref_filter(request.GET.get("consumo_ref"))
        consumo_limit = _parse_bounded_int(request.GET.get("limit", 30), default=30, min_value=1, max_value=1000)
        consumo_offset = _parse_bounded_int(request.GET.get("offset", 0), default=0, min_value=0, max_value=200000)
        consumo_sort_by = (request.GET.get("sort_by") or "variacion_cost_abs").strip().lower()
        consumo_sort_dir = (request.GET.get("sort_dir") or "desc").strip().lower()
        consumo_allowed_sort = {
            "variacion_cost_abs",
            "variacion_cost",
            "costo_real",
            "costo_plan",
            "cantidad_real",
            "cantidad_plan",
            "consumo_pct",
            "insumo",
            "categoria",
            "estado",
            "semaforo",
        }
        if consumo_sort_by not in consumo_allowed_sort:
            return Response(
                {
                    "detail": (
                        "sort_by inválido. Usa variacion_cost_abs, variacion_cost, costo_real, costo_plan, "
                        "cantidad_real, cantidad_plan, consumo_pct, insumo, categoria, estado o semaforo."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        if consumo_sort_dir not in {"asc", "desc"}:
            return Response(
                {"detail": "sort_dir inválido. Usa asc o desc."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        budget_ctx = _build_budget_context(
            solicitudes,
            source_filter,
            plan_filter,
            categoria_filter,
            _periodo_tipo,
            _periodo_mes,
        )
        provider_dashboard = _build_provider_dashboard(
            _periodo_mes,
            source_filter,
            plan_filter,
            categoria_filter,
            budget_ctx["presupuesto_rows_proveedor"],
        )
        category_dashboard = _build_category_dashboard(
            _periodo_mes,
            source_filter,
            plan_filter,
            categoria_filter,
            budget_ctx.get("presupuesto_rows_categoria", []),
        )
        consumo_dashboard = _build_consumo_vs_plan_dashboard(
            _periodo_tipo,
            _periodo_mes,
            source_filter,
            plan_filter,
            categoria_filter,
            consumo_ref_filter,
            limit=consumo_limit,
            offset=consumo_offset,
            sort_by=consumo_sort_by,
            sort_dir=consumo_sort_dir,
        )
        historial = _build_budget_history(_periodo_mes, source_filter, plan_filter, categoria_filter)

        payload = {
            "periodo": {
                "path": periodo,
                "tipo": _periodo_tipo,
                "mes": _periodo_mes,
                "label": periodo_label,
            },
            "filters": {
                "source": source_filter,
                "plan_id": plan_filter or "",
                "categoria": categoria_filter or "",
                "reabasto": reabasto_filter,
                "consumo_ref": consumo_ref_filter,
                "limit": consumo_limit,
                "offset": consumo_offset,
                "sort_by": consumo_sort_by,
                "sort_dir": consumo_sort_dir,
            },
            "totals": {
                "solicitudes_count": len(solicitudes),
                "presupuesto_estimado_total": _to_float(budget_ctx["presupuesto_estimado_total"]),
                "presupuesto_ejecutado_total": _to_float(budget_ctx["presupuesto_ejecutado_total"]),
                "presupuesto_objetivo": _to_float(budget_ctx.get("presupuesto_objetivo")),
                "presupuesto_variacion_objetivo": _to_float(budget_ctx.get("presupuesto_variacion_objetivo")),
                "alertas_total": int(budget_ctx.get("presupuesto_alertas_total") or 0),
                "alertas_excedidas": int(budget_ctx.get("presupuesto_alertas_excedidas") or 0),
                "alertas_preventivas": int(budget_ctx.get("presupuesto_alertas_preventivas") or 0),
            },
            "alertas": [
                {
                    "nivel": row.get("nivel"),
                    "scope": row.get("scope"),
                    "nombre": row.get("nombre"),
                    "estimado": _to_float(row.get("estimado")),
                    "ejecutado": _to_float(row.get("ejecutado")),
                    "objetivo": _to_float(row.get("objetivo")),
                    "uso_objetivo_pct": _to_float(row.get("uso_objetivo_pct")) if row.get("uso_objetivo_pct") is not None else None,
                    "variacion": _to_float(row.get("variacion")),
                    "estado": row.get("estado"),
                }
                for row in budget_ctx.get("presupuesto_alertas", [])[:100]
            ],
            "proveedores": [
                {
                    "proveedor": row["proveedor"],
                    "estimado": _to_float(row["estimado"]),
                    "ejecutado": _to_float(row["ejecutado"]),
                    "variacion": _to_float(row["variacion"]),
                    "participacion_pct": _to_float(row["participacion_pct"]),
                    "objetivo_proveedor": _to_float(row.get("objetivo_proveedor")),
                    "uso_objetivo_pct": _to_float(row.get("uso_objetivo_pct")) if row.get("uso_objetivo_pct") is not None else None,
                    "objetivo_estado": row.get("objetivo_estado"),
                }
                for row in budget_ctx["presupuesto_rows_proveedor"][:50]
            ],
            "categorias": [
                {
                    "categoria": row["categoria"],
                    "estimado": _to_float(row["estimado"]),
                    "ejecutado": _to_float(row["ejecutado"]),
                    "variacion": _to_float(row["variacion"]),
                    "participacion_pct": _to_float(row["participacion_pct"]),
                    "objetivo_categoria": _to_float(row.get("objetivo_categoria")),
                    "uso_objetivo_pct": _to_float(row.get("uso_objetivo_pct")) if row.get("uso_objetivo_pct") is not None else None,
                    "objetivo_estado": row.get("objetivo_estado"),
                }
                for row in budget_ctx.get("presupuesto_rows_categoria", [])[:50]
            ],
            "historial_6m": [
                {
                    "periodo_mes": row["periodo_mes"],
                    "objetivo": _to_float(row["objetivo"]),
                    "estimado": _to_float(row["estimado"]),
                    "ejecutado": _to_float(row["ejecutado"]),
                    "ratio_pct": _to_float(row["ratio_pct"]) if row.get("ratio_pct") is not None else None,
                    "estado_label": row["estado_label"],
                }
                for row in historial
            ],
            "trend": {
                "proveedor_rows": [
                    {
                        "proveedor": row["proveedor"],
                        "mes": row["mes"],
                        "estimado": _to_float(row["estimado"]),
                        "ejecutado": _to_float(row["ejecutado"]),
                        "variacion": _to_float(row["variacion"]),
                    }
                    for row in provider_dashboard["trend_rows"]
                ],
                "categoria_rows": [
                    {
                        "categoria": row["categoria"],
                        "mes": row["mes"],
                        "estimado": _to_float(row["estimado"]),
                        "ejecutado": _to_float(row["ejecutado"]),
                        "variacion": _to_float(row["variacion"]),
                    }
                    for row in category_dashboard["trend_rows"]
                ],
            },
            "consumo_vs_plan": {
                "meta": {
                    "rows_total": int(consumo_dashboard.get("meta", {}).get("rows_total") or 0),
                    "rows_returned": int(consumo_dashboard.get("meta", {}).get("rows_returned") or 0),
                    "limit": int(consumo_dashboard.get("meta", {}).get("limit") or consumo_limit),
                    "offset": int(consumo_dashboard.get("meta", {}).get("offset") or consumo_offset),
                    "sort_by": consumo_dashboard.get("meta", {}).get("sort_by") or consumo_sort_by,
                    "sort_dir": consumo_dashboard.get("meta", {}).get("sort_dir") or consumo_sort_dir,
                },
                "totals": {
                    "plan_qty_total": _to_float(consumo_dashboard["totals"]["plan_qty_total"]),
                    "consumo_real_qty_total": _to_float(consumo_dashboard["totals"]["consumo_real_qty_total"]),
                    "plan_cost_total": _to_float(consumo_dashboard["totals"]["plan_cost_total"]),
                    "consumo_real_cost_total": _to_float(consumo_dashboard["totals"]["consumo_real_cost_total"]),
                    "variacion_cost_total": _to_float(consumo_dashboard["totals"]["variacion_cost_total"]),
                    "semaforo_rojo_count": int(consumo_dashboard["totals"]["semaforo_rojo_count"] or 0),
                    "semaforo_amarillo_count": int(consumo_dashboard["totals"]["semaforo_amarillo_count"] or 0),
                    "semaforo_verde_count": int(consumo_dashboard["totals"]["semaforo_verde_count"] or 0),
                    "sin_costo_count": int(consumo_dashboard["totals"]["sin_costo_count"] or 0),
                    "cobertura_pct": (
                        _to_float(consumo_dashboard["totals"]["cobertura_pct"])
                        if consumo_dashboard["totals"]["cobertura_pct"] is not None
                        else None
                    ),
                },
                "rows": [
                    {
                        "insumo_id": row["insumo_id"],
                        "insumo": row["insumo"],
                        "categoria": row["categoria"],
                        "unidad": row["unidad"],
                        "cantidad_plan": _to_float(row["cantidad_plan"]),
                        "cantidad_real": _to_float(row["cantidad_real"]),
                        "variacion_qty": _to_float(row["variacion_qty"]),
                        "costo_unitario": _to_float(row["costo_unitario"]),
                        "costo_plan": _to_float(row["costo_plan"]),
                        "costo_real": _to_float(row["costo_real"]),
                        "variacion_cost": _to_float(row["variacion_cost"]),
                        "consumo_pct": _to_float(row["consumo_pct"]) if row["consumo_pct"] is not None else None,
                        "estado": row["estado"],
                        "semaforo": row["semaforo"],
                        "sin_costo": bool(row["sin_costo"]),
                        "alerta": row["alerta"],
                    }
                    for row in consumo_dashboard["rows"]
                ],
            },
        }

        return Response(payload, status=status.HTTP_200_OK)


class ComprasSolicitudesImportPreviewView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not can_manage_compras(request.user):
            return Response(
                {"detail": "No tienes permisos para importar solicitudes de compras."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = ComprasSolicitudImportPreviewSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        periodo_tipo = (data.get("periodo_tipo") or "mes").strip() or "mes"
        periodo_mes = (data.get("periodo_mes") or "").strip()
        default_fecha = _default_fecha_requerida(periodo_tipo, periodo_mes)
        area_default = (data.get("area_default") or "Compras").strip() or "Compras"
        solicitante_default = (data.get("solicitante_default") or request.user.username).strip() or request.user.username
        estatus_default = (data.get("estatus_default") or SolicitudCompra.STATUS_BORRADOR).strip().upper()
        valid_status = {x[0] for x in SolicitudCompra.STATUS_CHOICES}
        if estatus_default not in valid_status:
            estatus_default = SolicitudCompra.STATUS_BORRADOR
        evitar_duplicados = bool(data.get("evitar_duplicados", True))
        min_score = int(data.get("score_min") or 90)
        top = int(data.get("top") or 120)

        provider_map = {
            normalizar_nombre(p.nombre): p
            for p in Proveedor.objects.filter(activo=True).only("id", "nombre")
        }
        match_cache: dict[str, tuple[Insumo | None, float, str]] = {}
        parsed_rows: list[dict] = []
        duplicate_keys_to_check: set[tuple[str, int, date]] = set()

        for idx, row in enumerate(data.get("rows") or [], start=1):
            insumo_raw = str(row.get("insumo") or "").strip()
            cantidad = _to_decimal(row.get("cantidad"), Decimal("0"))
            if insumo_raw:
                cache_key = normalizar_nombre(insumo_raw)
                insumo_match, score, method = match_cache.get(cache_key, (None, 0.0, "sin_match"))
                if cache_key not in match_cache:
                    insumo_match, score, method = match_insumo(insumo_raw)
                    match_cache[cache_key] = (insumo_match, score, method)
            else:
                insumo_match, score, method = (None, 0.0, "sin_match")
            insumo_id = int(insumo_match.id) if (insumo_match and score >= min_score) else 0

            area = str(row.get("area") or area_default).strip() or area_default
            solicitante = str(row.get("solicitante") or solicitante_default).strip() or solicitante_default
            fecha_requerida = _parse_date_value(row.get("fecha_requerida"), default_fecha)
            estatus = str(row.get("estatus") or estatus_default).strip().upper()
            if estatus not in valid_status:
                estatus = estatus_default
            proveedor = _resolve_proveedor_name(str(row.get("proveedor") or ""), provider_map)
            if not proveedor and insumo_match:
                proveedor = insumo_match.proveedor_principal

            parsed_rows.append(
                {
                    "row_id": str(idx),
                    "source_row": idx,
                    "insumo_origen": insumo_raw,
                    "insumo_sugerencia": insumo_match.nombre if insumo_match else "",
                    "insumo_id": insumo_id,
                    "cantidad": cantidad,
                    "area": area,
                    "solicitante": solicitante,
                    "fecha_requerida": fecha_requerida,
                    "estatus": estatus,
                    "proveedor_id": int(proveedor.id) if proveedor else 0,
                    "score": float(score or 0),
                    "metodo": method,
                    "has_insumo_match": bool(insumo_match),
                }
            )
            if evitar_duplicados and insumo_id:
                duplicate_keys_to_check.add((area, insumo_id, fecha_requerida))

        duplicates_found: set[tuple[str, int, date]] = set()
        if duplicate_keys_to_check:
            areas = sorted({k[0] for k in duplicate_keys_to_check})
            insumo_ids = sorted({k[1] for k in duplicate_keys_to_check})
            fechas = sorted({k[2] for k in duplicate_keys_to_check})
            duplicates_found = {
                (area, int(insumo_id), fecha)
                for area, insumo_id, fecha in SolicitudCompra.objects.filter(
                    area__in=areas,
                    insumo_id__in=insumo_ids,
                    fecha_requerida__in=fechas,
                    estatus__in=_active_solicitud_statuses(),
                ).values_list("area", "insumo_id", "fecha_requerida")
            }

        preview_cost_by_insumo: dict[int, Decimal] = {}
        preview_insumo_ids = sorted({int(p["insumo_id"]) for p in parsed_rows if int(p["insumo_id"] or 0) > 0})
        if preview_insumo_ids:
            for c in CostoInsumo.objects.filter(insumo_id__in=preview_insumo_ids).order_by("insumo_id", "-fecha", "-id"):
                if c.insumo_id not in preview_cost_by_insumo:
                    preview_cost_by_insumo[c.insumo_id] = c.costo_unitario

        preview_rows: list[dict] = []
        ready_count = 0
        duplicates_count = 0
        without_match_count = 0
        invalid_qty_count = 0
        ready_qty_total = Decimal("0")
        ready_budget_total = Decimal("0")
        for parsed in parsed_rows:
            insumo_raw = str(parsed["insumo_origen"] or "").strip()
            cantidad = parsed["cantidad"]
            insumo_id = parsed["insumo_id"]
            costo_unitario = preview_cost_by_insumo.get(insumo_id, Decimal("0")) if insumo_id else Decimal("0")
            presupuesto_estimado = (cantidad * costo_unitario) if cantidad > 0 else Decimal("0")
            area = parsed["area"]
            fecha_requerida = parsed["fecha_requerida"]
            duplicate = bool(insumo_id and (area, insumo_id, fecha_requerida) in duplicates_found)

            notes: list[str] = []
            hard_error = False
            if not insumo_raw:
                notes.append("Insumo vacío en fila.")
                hard_error = True
            if not insumo_id:
                if parsed["has_insumo_match"]:
                    notes.append(f"Score de match insuficiente (<{min_score}).")
                else:
                    notes.append("Sin match de insumo.")
                hard_error = True
                without_match_count += 1
            if cantidad <= 0:
                notes.append("Cantidad inválida (debe ser > 0).")
                hard_error = True
                invalid_qty_count += 1
            if duplicate:
                notes.append("Posible duplicado con solicitud activa.")
                duplicates_count += 1

            include = not hard_error
            if include and cantidad > 0:
                ready_count += 1
                ready_qty_total += cantidad
                ready_budget_total += max(presupuesto_estimado, Decimal("0"))

            preview_rows.append(
                {
                    "row_id": parsed["row_id"],
                    "source_row": parsed["source_row"],
                    "insumo_origen": insumo_raw,
                    "insumo_sugerencia": parsed["insumo_sugerencia"],
                    "insumo_id": parsed["insumo_id"] or None,
                    "cantidad": str(cantidad),
                    "area": parsed["area"],
                    "solicitante": parsed["solicitante"],
                    "fecha_requerida": fecha_requerida.isoformat(),
                    "estatus": parsed["estatus"],
                    "proveedor_id": parsed["proveedor_id"] or None,
                    "score": round(float(parsed["score"] or 0), 1),
                    "metodo": parsed["metodo"],
                    "costo_unitario": str(costo_unitario),
                    "presupuesto_estimado": str(presupuesto_estimado),
                    "duplicate": duplicate,
                    "notes": " | ".join(notes),
                    "include": include,
                }
            )

        return Response(
            {
                "preview": {
                    "periodo_tipo": periodo_tipo,
                    "periodo_mes": periodo_mes,
                    "evitar_duplicados": evitar_duplicados,
                    "score_min": min_score,
                    "rows": preview_rows[:top],
                },
                "totales": {
                    "filas": len(preview_rows),
                    "ready_count": ready_count,
                    "excluded_count": max(0, len(preview_rows) - ready_count),
                    "duplicates_count": duplicates_count,
                    "without_match_count": without_match_count,
                    "invalid_qty_count": invalid_qty_count,
                    "ready_qty_total": str(ready_qty_total),
                    "ready_budget_total": str(ready_budget_total),
                },
            },
            status=status.HTTP_200_OK,
        )


class ComprasSolicitudesImportConfirmView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not can_manage_compras(request.user):
            return Response(
                {"detail": "No tienes permisos para confirmar importación de solicitudes."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = ComprasSolicitudImportConfirmSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        rows = data.get("rows") or []
        periodo_tipo = (data.get("periodo_tipo") or "mes").strip() or "mes"
        periodo_mes = (data.get("periodo_mes") or "").strip()
        evitar_duplicados = bool(data.get("evitar_duplicados", True))
        default_fecha = _default_fecha_requerida(periodo_tipo, periodo_mes)
        valid_status = {x[0] for x in SolicitudCompra.STATUS_CHOICES}

        included_rows = [row for row in rows if bool(row.get("include", True))]
        insumo_ids = sorted({int(row.get("insumo_id") or 0) for row in included_rows if int(row.get("insumo_id") or 0) > 0})
        proveedor_ids = sorted(
            {
                int(row.get("proveedor_id") or 0)
                for row in included_rows
                if int(row.get("proveedor_id") or 0) > 0
            }
        )

        insumos_map = Insumo.objects.select_related("proveedor_principal").in_bulk(insumo_ids)
        proveedores_map = Proveedor.objects.filter(activo=True, id__in=proveedor_ids).in_bulk()

        existing_duplicate_keys: set[tuple[str, int, date]] = set()
        if evitar_duplicados:
            batch_keys: set[tuple[str, int, date]] = set()
            for row in included_rows:
                area = (str(row.get("area") or "").strip() or "General")
                insumo_id = int(row.get("insumo_id") or 0)
                if insumo_id <= 0:
                    continue
                fecha_requerida = _parse_date_value(row.get("fecha_requerida"), default_fecha)
                batch_keys.add((area, insumo_id, fecha_requerida))
            if batch_keys:
                areas = sorted({k[0] for k in batch_keys})
                insumo_ids_batch = sorted({k[1] for k in batch_keys})
                fechas = sorted({k[2] for k in batch_keys})
                existing_duplicate_keys = {
                    (area, int(insumo_id), fecha)
                    for area, insumo_id, fecha in SolicitudCompra.objects.filter(
                        area__in=areas,
                        insumo_id__in=insumo_ids_batch,
                        fecha_requerida__in=fechas,
                        estatus__in=_active_solicitud_statuses(),
                    ).values_list("area", "insumo_id", "fecha_requerida")
                }

        created = 0
        skipped_invalid = 0
        skipped_duplicate = 0
        skipped_removed = max(0, len(rows) - len(included_rows))
        created_duplicate_keys: set[tuple[str, int, date]] = set()
        created_items: list[dict] = []

        with transaction.atomic():
            for row in rows:
                if not bool(row.get("include", True)):
                    continue

                area = (str(row.get("area") or "").strip() or "General")
                solicitante = (str(row.get("solicitante") or "").strip() or request.user.username)
                estatus = (str(row.get("estatus") or SolicitudCompra.STATUS_BORRADOR).strip().upper())
                if estatus not in valid_status:
                    estatus = SolicitudCompra.STATUS_BORRADOR

                insumo_id = int(row.get("insumo_id") or 0)
                insumo = insumos_map.get(insumo_id)
                if not insumo:
                    skipped_invalid += 1
                    continue

                cantidad = _to_decimal(row.get("cantidad"), Decimal("0"))
                if cantidad <= 0:
                    skipped_invalid += 1
                    continue

                fecha_requerida = _parse_date_value(row.get("fecha_requerida"), default_fecha)
                proveedor = proveedores_map.get(int(row.get("proveedor_id") or 0))
                if not proveedor:
                    proveedor = insumo.proveedor_principal

                duplicate_key = (area, int(insumo.id), fecha_requerida)
                if evitar_duplicados and ((duplicate_key in existing_duplicate_keys) or (duplicate_key in created_duplicate_keys)):
                    skipped_duplicate += 1
                    continue

                solicitud = SolicitudCompra.objects.create(
                    area=area,
                    solicitante=solicitante,
                    insumo=insumo,
                    proveedor_sugerido=proveedor,
                    cantidad=cantidad,
                    fecha_requerida=fecha_requerida,
                    estatus=estatus,
                )
                log_event(
                    request.user,
                    "CREATE",
                    "compras.SolicitudCompra",
                    solicitud.id,
                    {"folio": solicitud.folio, "source": "api_import_confirm"},
                )
                if evitar_duplicados:
                    created_duplicate_keys.add(duplicate_key)
                created += 1
                if len(created_items) < 80:
                    created_items.append(
                        {
                            "id": solicitud.id,
                            "folio": solicitud.folio,
                            "insumo": solicitud.insumo.nombre,
                            "cantidad": str(solicitud.cantidad),
                            "fecha_requerida": solicitud.fecha_requerida.isoformat(),
                            "estatus": solicitud.estatus,
                        }
                    )

        return Response(
            {
                "periodo_tipo": periodo_tipo,
                "periodo_mes": periodo_mes,
                "totales": {
                    "rows": len(rows),
                    "created": created,
                    "skipped_removed": skipped_removed,
                    "skipped_duplicate": skipped_duplicate,
                    "skipped_invalid": skipped_invalid,
                },
                "items": created_items,
            },
            status=status.HTTP_200_OK,
        )


class ComprasSolicitudesListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not can_view_compras(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar solicitudes de compra."},
                status=status.HTTP_403_FORBIDDEN,
            )

        estatus = (request.GET.get("estatus") or "").strip().upper()
        area = (request.GET.get("area") or "").strip()
        q = (request.GET.get("q") or "").strip()
        periodo = (request.GET.get("mes") or request.GET.get("periodo") or "").strip()
        limit = _parse_bounded_int(request.GET.get("limit", 120), default=120, min_value=1, max_value=500)
        offset = _parse_bounded_int(request.GET.get("offset", 0), default=0, min_value=0, max_value=200000)
        sort_by = (request.GET.get("sort_by") or "creado_en").strip().lower()
        sort_dir = (request.GET.get("sort_dir") or "desc").strip().lower()
        allowed_sort = {
            "creado_en": "creado_en",
            "folio": "folio",
            "fecha_requerida": "fecha_requerida",
            "cantidad": "cantidad",
            "estatus": "estatus",
            "area": "area",
            "solicitante": "solicitante",
            "insumo": "insumo__nombre",
            "proveedor": "proveedor_sugerido__nombre",
        }
        if sort_by not in allowed_sort:
            return Response(
                {
                    "detail": (
                        "sort_by inválido. Usa creado_en, folio, fecha_requerida, cantidad, estatus, "
                        "area, solicitante, insumo o proveedor."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        if sort_dir not in {"asc", "desc"}:
            return Response(
                {"detail": "sort_dir inválido. Usa asc o desc."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = (
            SolicitudCompra.objects.select_related("insumo", "insumo__unidad_base", "proveedor_sugerido")
            .order_by("-creado_en")
        )
        valid_status = {choice[0] for choice in SolicitudCompra.STATUS_CHOICES}
        if estatus in valid_status:
            qs = qs.filter(estatus=estatus)
        else:
            estatus = ""
        if area:
            qs = qs.filter(area__icontains=area)
        parsed_period = _parse_period(periodo)
        if parsed_period:
            y, m = parsed_period
            qs = qs.filter(fecha_requerida__year=y, fecha_requerida__month=m)
            periodo = f"{y:04d}-{m:02d}"
        else:
            periodo = ""
        if q:
            qs = qs.filter(
                Q(folio__icontains=q)
                | Q(solicitante__icontains=q)
                | Q(area__icontains=q)
                | Q(insumo__nombre__icontains=q)
                | Q(proveedor_sugerido__nombre__icontains=q)
            )
        sort_field = allowed_sort[sort_by]
        order_expr = sort_field if sort_dir == "asc" else f"-{sort_field}"
        qs = qs.order_by(order_expr, "-id")

        rows_total = qs.count()
        by_status = {k: 0 for k in valid_status}
        for row in qs.order_by().values("estatus").annotate(rows=Count("id")):
            key = row.get("estatus")
            if key in by_status:
                by_status[key] = int(row.get("rows") or 0)

        insumo_totals = list(qs.order_by().values("insumo_id").annotate(total_qty=Sum("cantidad")))
        rows = list(qs[offset : offset + limit])
        insumo_ids = [r.insumo_id for r in rows]
        insumo_ids_totales = [int(x["insumo_id"]) for x in insumo_totals if x.get("insumo_id")]
        costo_ids = sorted(set(insumo_ids) | set(insumo_ids_totales))
        latest_cost_by_insumo: dict[int, Decimal] = {}
        if costo_ids:
            for c in CostoInsumo.objects.filter(insumo_id__in=costo_ids).order_by("insumo_id", "-fecha", "-id"):
                if c.insumo_id not in latest_cost_by_insumo:
                    latest_cost_by_insumo[c.insumo_id] = _to_decimal(c.costo_unitario)

        items = []
        presupuesto_total = Decimal("0")
        presupuesto_total_filtered = Decimal("0")
        for row in insumo_totals:
            insumo_id = int(row.get("insumo_id") or 0)
            qty = _to_decimal(row.get("total_qty"))
            costo_unitario = _to_decimal(latest_cost_by_insumo.get(insumo_id, 0))
            presupuesto_total_filtered += (qty * costo_unitario).quantize(Decimal("0.01"))
        for r in rows:
            costo_unitario = _to_decimal(latest_cost_by_insumo.get(r.insumo_id, 0))
            presupuesto = (_to_decimal(r.cantidad) * costo_unitario).quantize(Decimal("0.01"))
            presupuesto_total += presupuesto
            items.append(
                {
                    "id": r.id,
                    "folio": r.folio,
                    "estatus": r.estatus,
                    "area": r.area,
                    "solicitante": r.solicitante,
                    "insumo_id": r.insumo_id,
                    "insumo": r.insumo.nombre,
                    "unidad": r.insumo.unidad_base.codigo if r.insumo.unidad_base_id and r.insumo.unidad_base else "",
                    "cantidad": str(r.cantidad),
                    "fecha_requerida": str(r.fecha_requerida),
                    "proveedor_sugerido": r.proveedor_sugerido.nombre if r.proveedor_sugerido_id else "",
                    "costo_unitario": str(costo_unitario),
                    "presupuesto_estimado": str(presupuesto),
                    "creado_en": r.creado_en,
                }
            )

        return Response(
            {
                "filters": {
                    "estatus": estatus,
                    "area": area,
                    "q": q,
                    "periodo": periodo,
                    "limit": limit,
                    "offset": offset,
                    "sort_by": sort_by,
                    "sort_dir": sort_dir,
                },
                "totales": {
                    "rows": len(items),
                    "rows_total": rows_total,
                    "rows_returned": len(items),
                    "presupuesto_estimado_total": str(presupuesto_total),
                    "presupuesto_estimado_total_filtered": str(presupuesto_total_filtered),
                    "by_status": by_status,
                },
                "items": items,
            },
            status=status.HTTP_200_OK,
        )


class ComprasOrdenesListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not can_view_compras(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar órdenes de compra."},
                status=status.HTTP_403_FORBIDDEN,
            )

        estatus = (request.GET.get("estatus") or "").strip().upper()
        q = (request.GET.get("q") or "").strip()
        periodo = (request.GET.get("mes") or request.GET.get("periodo") or "").strip()
        proveedor_id_raw = (request.GET.get("proveedor_id") or "").strip()
        limit = _parse_bounded_int(request.GET.get("limit", 120), default=120, min_value=1, max_value=500)
        offset = _parse_bounded_int(request.GET.get("offset", 0), default=0, min_value=0, max_value=200000)
        sort_by = (request.GET.get("sort_by") or "creado_en").strip().lower()
        sort_dir = (request.GET.get("sort_dir") or "desc").strip().lower()
        allowed_sort = {
            "creado_en": "creado_en",
            "folio": "folio",
            "fecha_emision": "fecha_emision",
            "fecha_entrega_estimada": "fecha_entrega_estimada",
            "monto_estimado": "monto_estimado",
            "estatus": "estatus",
            "proveedor": "proveedor__nombre",
            "referencia": "referencia",
        }
        if sort_by not in allowed_sort:
            return Response(
                {
                    "detail": (
                        "sort_by inválido. Usa creado_en, folio, fecha_emision, fecha_entrega_estimada, "
                        "monto_estimado, estatus, proveedor o referencia."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        if sort_dir not in {"asc", "desc"}:
            return Response(
                {"detail": "sort_dir inválido. Usa asc o desc."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = OrdenCompra.objects.select_related("proveedor", "solicitud", "solicitud__insumo").order_by("-creado_en")
        valid_status = {choice[0] for choice in OrdenCompra.STATUS_CHOICES}
        if estatus in valid_status:
            qs = qs.filter(estatus=estatus)
        else:
            estatus = ""

        if proveedor_id_raw.isdigit():
            qs = qs.filter(proveedor_id=int(proveedor_id_raw))
        else:
            proveedor_id_raw = ""

        parsed_period = _parse_period(periodo)
        if parsed_period:
            y, m = parsed_period
            qs = qs.filter(fecha_emision__year=y, fecha_emision__month=m)
            periodo = f"{y:04d}-{m:02d}"
        else:
            periodo = ""

        if q:
            qs = qs.filter(
                Q(folio__icontains=q)
                | Q(referencia__icontains=q)
                | Q(proveedor__nombre__icontains=q)
                | Q(solicitud__folio__icontains=q)
            )
        sort_field = allowed_sort[sort_by]
        order_expr = sort_field if sort_dir == "asc" else f"-{sort_field}"
        qs = qs.order_by(order_expr, "-id")

        rows_total = qs.count()
        monto_total_filtered = _to_decimal(qs.aggregate(total=Sum("monto_estimado"))["total"])
        by_status = {k: 0 for k in valid_status}
        for row in qs.order_by().values("estatus").annotate(rows=Count("id")):
            key = row.get("estatus")
            if key in by_status:
                by_status[key] = int(row.get("rows") or 0)

        rows = list(qs[offset : offset + limit])
        items = []
        monto_total = Decimal("0")
        for r in rows:
            monto = _to_decimal(r.monto_estimado)
            monto_total += monto
            items.append(
                {
                    "id": r.id,
                    "folio": r.folio,
                    "estatus": r.estatus,
                    "proveedor_id": r.proveedor_id,
                    "proveedor": r.proveedor.nombre if r.proveedor_id else "",
                    "solicitud_id": r.solicitud_id,
                    "solicitud_folio": r.solicitud.folio if r.solicitud_id and r.solicitud else "",
                    "insumo": (
                        r.solicitud.insumo.nombre
                        if r.solicitud_id and getattr(r.solicitud, "insumo_id", None) and r.solicitud.insumo
                        else ""
                    ),
                    "referencia": r.referencia,
                    "fecha_emision": str(r.fecha_emision) if r.fecha_emision else "",
                    "fecha_entrega_estimada": str(r.fecha_entrega_estimada) if r.fecha_entrega_estimada else "",
                    "monto_estimado": str(monto),
                    "creado_en": r.creado_en,
                }
            )

        return Response(
            {
                "filters": {
                    "estatus": estatus,
                    "q": q,
                    "periodo": periodo,
                    "proveedor_id": proveedor_id_raw,
                    "limit": limit,
                    "offset": offset,
                    "sort_by": sort_by,
                    "sort_dir": sort_dir,
                },
                "totales": {
                    "rows": len(items),
                    "rows_total": rows_total,
                    "rows_returned": len(items),
                    "monto_estimado_total": str(monto_total),
                    "monto_estimado_total_filtered": str(monto_total_filtered),
                    "by_status": by_status,
                },
                "items": items,
            },
            status=status.HTTP_200_OK,
        )


class ComprasRecepcionesListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not can_view_compras(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar recepciones de compra."},
                status=status.HTTP_403_FORBIDDEN,
            )

        estatus = (request.GET.get("estatus") or "").strip().upper()
        q = (request.GET.get("q") or "").strip()
        periodo = (request.GET.get("mes") or request.GET.get("periodo") or "").strip()
        proveedor_id_raw = (request.GET.get("proveedor_id") or "").strip()
        limit = _parse_bounded_int(request.GET.get("limit", 120), default=120, min_value=1, max_value=500)
        offset = _parse_bounded_int(request.GET.get("offset", 0), default=0, min_value=0, max_value=200000)
        sort_by = (request.GET.get("sort_by") or "creado_en").strip().lower()
        sort_dir = (request.GET.get("sort_dir") or "desc").strip().lower()
        allowed_sort = {
            "creado_en": "creado_en",
            "folio": "folio",
            "fecha_recepcion": "fecha_recepcion",
            "conformidad_pct": "conformidad_pct",
            "estatus": "estatus",
            "proveedor": "orden__proveedor__nombre",
            "orden_folio": "orden__folio",
        }
        if sort_by not in allowed_sort:
            return Response(
                {
                    "detail": (
                        "sort_by inválido. Usa creado_en, folio, fecha_recepcion, conformidad_pct, "
                        "estatus, proveedor u orden_folio."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        if sort_dir not in {"asc", "desc"}:
            return Response(
                {"detail": "sort_dir inválido. Usa asc o desc."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = RecepcionCompra.objects.select_related("orden", "orden__proveedor").order_by("-creado_en")
        valid_status = {choice[0] for choice in RecepcionCompra.STATUS_CHOICES}
        if estatus in valid_status:
            qs = qs.filter(estatus=estatus)
        else:
            estatus = ""

        if proveedor_id_raw.isdigit():
            qs = qs.filter(orden__proveedor_id=int(proveedor_id_raw))
        else:
            proveedor_id_raw = ""

        parsed_period = _parse_period(periodo)
        if parsed_period:
            y, m = parsed_period
            qs = qs.filter(fecha_recepcion__year=y, fecha_recepcion__month=m)
            periodo = f"{y:04d}-{m:02d}"
        else:
            periodo = ""

        if q:
            qs = qs.filter(
                Q(folio__icontains=q)
                | Q(orden__folio__icontains=q)
                | Q(orden__proveedor__nombre__icontains=q)
                | Q(observaciones__icontains=q)
            )
        sort_field = allowed_sort[sort_by]
        order_expr = sort_field if sort_dir == "asc" else f"-{sort_field}"
        qs = qs.order_by(order_expr, "-id")

        rows_total = qs.count()
        by_status = {k: 0 for k in valid_status}
        for row in qs.order_by().values("estatus").annotate(rows=Count("id")):
            key = row.get("estatus")
            if key in by_status:
                by_status[key] = int(row.get("rows") or 0)

        rows = list(qs[offset : offset + limit])
        items = []
        for r in rows:
            items.append(
                {
                    "id": r.id,
                    "folio": r.folio,
                    "estatus": r.estatus,
                    "orden_id": r.orden_id,
                    "orden_folio": r.orden.folio if r.orden_id and r.orden else "",
                    "orden_estatus": r.orden.estatus if r.orden_id and r.orden else "",
                    "proveedor_id": r.orden.proveedor_id if r.orden_id and r.orden else None,
                    "proveedor": (
                        r.orden.proveedor.nombre
                        if r.orden_id and r.orden and r.orden.proveedor_id and r.orden.proveedor
                        else ""
                    ),
                    "fecha_recepcion": str(r.fecha_recepcion) if r.fecha_recepcion else "",
                    "conformidad_pct": str(_to_decimal(r.conformidad_pct)),
                    "observaciones": r.observaciones,
                    "creado_en": r.creado_en,
                }
            )

        return Response(
            {
                "filters": {
                    "estatus": estatus,
                    "q": q,
                    "periodo": periodo,
                    "proveedor_id": proveedor_id_raw,
                    "limit": limit,
                    "offset": offset,
                    "sort_by": sort_by,
                    "sort_dir": sort_dir,
                },
                "totales": {
                    "rows": len(items),
                    "rows_total": rows_total,
                    "rows_returned": len(items),
                    "by_status": by_status,
                },
                "items": items,
            },
            status=status.HTTP_200_OK,
        )


class ComprasSolicitudCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not can_manage_compras(request.user):
            return Response(
                {"detail": "No tienes permisos para crear solicitudes de compra."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = ComprasSolicitudCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        insumo = get_object_or_404(Insumo, pk=data["insumo_id"])
        solicitante = (data.get("solicitante") or request.user.username or "").strip() or request.user.username
        area = (data["area"] or "").strip() or "General"
        auto_crear_orden = bool(data.get("auto_crear_orden"))
        orden_estatus = data.get("orden_estatus") or OrdenCompra.STATUS_BORRADOR
        if auto_crear_orden and not insumo.proveedor_principal_id:
            return Response(
                {
                    "detail": (
                        "No se pudo crear OC automática: el insumo no tiene proveedor "
                        "principal configurado."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            solicitud = SolicitudCompra.objects.create(
                area=area,
                solicitante=solicitante[:120],
                insumo=insumo,
                proveedor_sugerido=insumo.proveedor_principal,
                cantidad=data["cantidad"],
                fecha_requerida=data.get("fecha_requerida") or timezone.localdate(),
                estatus=data.get("estatus") or SolicitudCompra.STATUS_BORRADOR,
            )

            orden = None
            if auto_crear_orden:
                proveedor = solicitud.proveedor_sugerido or insumo.proveedor_principal
                costo_qs = CostoInsumo.objects.filter(insumo=insumo).order_by("-fecha", "-id")
                costo = costo_qs.filter(proveedor=proveedor).first() or costo_qs.first()
                costo_unitario = _to_decimal(costo.costo_unitario if costo else 0)
                monto_estimado = (data["cantidad"] * costo_unitario).quantize(Decimal("0.01"))
                orden = OrdenCompra.objects.create(
                    solicitud=solicitud,
                    proveedor=proveedor,
                    estatus=orden_estatus,
                    monto_estimado=monto_estimado,
                )

        payload = {
            "id": solicitud.id,
            "folio": solicitud.folio,
            "area": solicitud.area,
            "solicitante": solicitud.solicitante,
            "insumo_id": solicitud.insumo_id,
            "insumo": solicitud.insumo.nombre,
            "cantidad": str(solicitud.cantidad),
            "fecha_requerida": str(solicitud.fecha_requerida),
            "estatus": solicitud.estatus,
            "proveedor_sugerido_id": solicitud.proveedor_sugerido_id,
            "proveedor_sugerido": solicitud.proveedor_sugerido.nombre if solicitud.proveedor_sugerido_id else "",
            "auto_crear_orden": auto_crear_orden,
            "orden_id": orden.id if orden else None,
            "orden_folio": orden.folio if orden else "",
            "orden_estatus": orden.estatus if orden else "",
        }
        return Response(payload, status=status.HTTP_201_CREATED)


class ComprasSolicitudStatusUpdateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, solicitud_id: int):
        if not can_manage_compras(request.user):
            return Response(
                {"detail": "No tienes permisos para aprobar/rechazar solicitudes de compra."},
                status=status.HTTP_403_FORBIDDEN,
            )

        solicitud = get_object_or_404(SolicitudCompra.objects.select_related("insumo"), pk=solicitud_id)
        ser = ComprasSolicitudStatusSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        estatus_new = ser.validated_data["estatus"]

        estatus_prev = solicitud.estatus
        if estatus_prev == estatus_new:
            return Response(
                {
                    "id": solicitud.id,
                    "folio": solicitud.folio,
                    "from": estatus_prev,
                    "to": estatus_new,
                    "updated": False,
                },
                status=status.HTTP_200_OK,
            )

        if not _can_transition_solicitud(estatus_prev, estatus_new):
            return Response(
                {
                    "detail": (
                        f"Transición inválida de solicitud: {estatus_prev} -> {estatus_new}."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        solicitud.estatus = estatus_new
        solicitud.save(update_fields=["estatus"])
        log_event(
            request.user,
            "APPROVE",
            "compras.SolicitudCompra",
            solicitud.id,
            {"from": estatus_prev, "to": estatus_new, "folio": solicitud.folio, "source": "api"},
        )

        return Response(
            {
                "id": solicitud.id,
                "folio": solicitud.folio,
                "from": estatus_prev,
                "to": estatus_new,
                "updated": True,
            },
            status=status.HTTP_200_OK,
        )


class ComprasSolicitudCrearOrdenView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, solicitud_id: int):
        if not can_manage_compras(request.user):
            return Response(
                {"detail": "No tienes permisos para crear órdenes desde solicitud."},
                status=status.HTTP_403_FORBIDDEN,
            )

        solicitud = get_object_or_404(
            SolicitudCompra.objects.select_related("insumo", "proveedor_sugerido", "insumo__proveedor_principal"),
            pk=solicitud_id,
        )
        if solicitud.estatus != SolicitudCompra.STATUS_APROBADA:
            return Response(
                {"detail": f"La solicitud {solicitud.folio} no está aprobada."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        existing = (
            OrdenCompra.objects.filter(solicitud=solicitud)
            .exclude(estatus=OrdenCompra.STATUS_CERRADA)
            .select_related("proveedor")
            .first()
        )
        if existing:
            return Response(
                {
                    "created": False,
                    "id": existing.id,
                    "folio": existing.folio,
                    "estatus": existing.estatus,
                    "proveedor": existing.proveedor.nombre,
                    "solicitud_folio": solicitud.folio,
                },
                status=status.HTTP_200_OK,
            )

        ser = ComprasCrearOrdenSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data
        estatus = data.get("estatus") or OrdenCompra.STATUS_BORRADOR
        if estatus not in {OrdenCompra.STATUS_BORRADOR, OrdenCompra.STATUS_ENVIADA}:
            return Response(
                {
                    "detail": (
                        "Estatus inicial de OC inválido. "
                        "Solo se permite BORRADOR o ENVIADA al crear desde solicitud."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        proveedor = None
        proveedor_id = data.get("proveedor_id")
        if proveedor_id:
            proveedor = get_object_or_404(Proveedor, pk=proveedor_id, activo=True)
        if proveedor is None:
            proveedor = solicitud.proveedor_sugerido or solicitud.insumo.proveedor_principal
        if proveedor is None:
            return Response(
                {
                    "detail": (
                        f"La solicitud {solicitud.folio} no tiene proveedor sugerido "
                        "ni proveedor principal en el insumo."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        costo_qs = CostoInsumo.objects.filter(insumo=solicitud.insumo).order_by("-fecha", "-id")
        costo = costo_qs.filter(proveedor=proveedor).first() or costo_qs.first()
        costo_unitario = _to_decimal(costo.costo_unitario if costo else 0)
        monto_estimado = ((solicitud.cantidad or Decimal("0")) * costo_unitario).quantize(Decimal("0.01"))

        orden = OrdenCompra.objects.create(
            solicitud=solicitud,
            proveedor=proveedor,
            referencia=f"SOLICITUD:{solicitud.folio}",
            fecha_emision=data.get("fecha_emision") or timezone.localdate(),
            fecha_entrega_estimada=data.get("fecha_entrega_estimada") or solicitud.fecha_requerida,
            monto_estimado=monto_estimado,
            estatus=estatus,
        )
        log_event(
            request.user,
            "CREATE",
            "compras.OrdenCompra",
            orden.id,
            {"folio": orden.folio, "estatus": orden.estatus, "source": f"api:solicitud:{solicitud.folio}"},
        )

        return Response(
            {
                "created": True,
                "id": orden.id,
                "folio": orden.folio,
                "estatus": orden.estatus,
                "solicitud_id": solicitud.id,
                "solicitud_folio": solicitud.folio,
                "proveedor_id": proveedor.id,
                "proveedor": proveedor.nombre,
                "monto_estimado": str(orden.monto_estimado),
                "costo_unitario": str(costo_unitario),
            },
            status=status.HTTP_201_CREATED,
        )


class ComprasOrdenStatusUpdateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, orden_id: int):
        if not can_manage_compras(request.user):
            return Response(
                {"detail": "No tienes permisos para operar órdenes de compra."},
                status=status.HTTP_403_FORBIDDEN,
            )

        orden = get_object_or_404(OrdenCompra, pk=orden_id)
        ser = ComprasOrdenStatusSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        estatus_new = ser.validated_data["estatus"]
        estatus_prev = orden.estatus

        if estatus_prev == estatus_new:
            return Response(
                {
                    "id": orden.id,
                    "folio": orden.folio,
                    "from": estatus_prev,
                    "to": estatus_new,
                    "updated": False,
                },
                status=status.HTTP_200_OK,
            )

        if estatus_new == OrdenCompra.STATUS_CERRADA:
            has_closed_recepcion = RecepcionCompra.objects.filter(
                orden=orden,
                estatus=RecepcionCompra.STATUS_CERRADA,
            ).exists()
            if not has_closed_recepcion:
                return Response(
                    {"detail": f"No puedes cerrar {orden.folio} sin al menos una recepción cerrada."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if not _can_transition_orden(estatus_prev, estatus_new):
            return Response(
                {"detail": f"Transición inválida de orden: {estatus_prev} -> {estatus_new}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        orden.estatus = estatus_new
        orden.save(update_fields=["estatus"])
        log_event(
            request.user,
            "APPROVE",
            "compras.OrdenCompra",
            orden.id,
            {"from": estatus_prev, "to": estatus_new, "folio": orden.folio, "source": "api"},
        )
        return Response(
            {
                "id": orden.id,
                "folio": orden.folio,
                "from": estatus_prev,
                "to": estatus_new,
                "updated": True,
            },
            status=status.HTTP_200_OK,
        )


class ComprasOrdenCreateRecepcionView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, orden_id: int):
        if not can_manage_compras(request.user):
            return Response(
                {"detail": "No tienes permisos para registrar recepciones."},
                status=status.HTTP_403_FORBIDDEN,
            )

        orden = get_object_or_404(OrdenCompra.objects.select_related("proveedor"), pk=orden_id)
        if orden.estatus in {OrdenCompra.STATUS_BORRADOR, OrdenCompra.STATUS_CERRADA}:
            return Response(
                {"detail": f"La orden {orden.folio} no admite recepciones en estatus {orden.estatus}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ser = ComprasRecepcionCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        recepcion = RecepcionCompra.objects.create(
            orden=orden,
            fecha_recepcion=data.get("fecha_recepcion") or timezone.localdate(),
            conformidad_pct=data.get("conformidad_pct", Decimal("100")),
            estatus=data.get("estatus") or RecepcionCompra.STATUS_PENDIENTE,
            observaciones=(data.get("observaciones") or "").strip(),
        )
        log_event(
            request.user,
            "CREATE",
            "compras.RecepcionCompra",
            recepcion.id,
            {"folio": recepcion.folio, "estatus": recepcion.estatus, "source": "api"},
        )

        if recepcion.estatus == RecepcionCompra.STATUS_CERRADA:
            _apply_recepcion_to_inventario(recepcion, acted_by=request.user)
            if orden.estatus != OrdenCompra.STATUS_CERRADA:
                orden_prev = orden.estatus
                orden.estatus = OrdenCompra.STATUS_CERRADA
                orden.save(update_fields=["estatus"])
                log_event(
                    request.user,
                    "APPROVE",
                    "compras.OrdenCompra",
                    orden.id,
                    {"from": orden_prev, "to": OrdenCompra.STATUS_CERRADA, "folio": orden.folio, "source": recepcion.folio},
                )

        return Response(
            {
                "id": recepcion.id,
                "folio": recepcion.folio,
                "orden_id": orden.id,
                "orden_folio": orden.folio,
                "estatus": recepcion.estatus,
                "conformidad_pct": str(recepcion.conformidad_pct),
                "fecha_recepcion": str(recepcion.fecha_recepcion),
                "orden_estatus": orden.estatus,
            },
            status=status.HTTP_201_CREATED,
        )


class ComprasRecepcionStatusUpdateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, recepcion_id: int):
        if not can_manage_compras(request.user):
            return Response(
                {"detail": "No tienes permisos para cerrar recepciones."},
                status=status.HTTP_403_FORBIDDEN,
            )

        recepcion = get_object_or_404(RecepcionCompra.objects.select_related("orden"), pk=recepcion_id)
        ser = ComprasRecepcionStatusSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        estatus_new = ser.validated_data["estatus"]
        estatus_prev = recepcion.estatus

        if estatus_prev == estatus_new:
            return Response(
                {
                    "id": recepcion.id,
                    "folio": recepcion.folio,
                    "from": estatus_prev,
                    "to": estatus_new,
                    "updated": False,
                    "orden_estatus": recepcion.orden.estatus,
                },
                status=status.HTTP_200_OK,
            )

        if not _can_transition_recepcion(estatus_prev, estatus_new):
            return Response(
                {"detail": f"Transición inválida de recepción: {estatus_prev} -> {estatus_new}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        recepcion.estatus = estatus_new
        recepcion.save(update_fields=["estatus"])
        log_event(
            request.user,
            "APPROVE",
            "compras.RecepcionCompra",
            recepcion.id,
            {"from": estatus_prev, "to": estatus_new, "folio": recepcion.folio, "source": "api"},
        )

        if estatus_new == RecepcionCompra.STATUS_CERRADA:
            _apply_recepcion_to_inventario(recepcion, acted_by=request.user)
            if recepcion.orden.estatus != OrdenCompra.STATUS_CERRADA:
                orden_prev = recepcion.orden.estatus
                recepcion.orden.estatus = OrdenCompra.STATUS_CERRADA
                recepcion.orden.save(update_fields=["estatus"])
                log_event(
                    request.user,
                    "APPROVE",
                    "compras.OrdenCompra",
                    recepcion.orden.id,
                    {"from": orden_prev, "to": OrdenCompra.STATUS_CERRADA, "folio": recepcion.orden.folio, "source": recepcion.folio},
                )

        return Response(
            {
                "id": recepcion.id,
                "folio": recepcion.folio,
                "from": estatus_prev,
                "to": estatus_new,
                "updated": True,
                "orden_id": recepcion.orden_id,
                "orden_folio": recepcion.orden.folio,
                "orden_estatus": recepcion.orden.estatus,
            },
            status=status.HTTP_200_OK,
        )


class MRPRequerimientosView(APIView):
    permission_classes = [IsAuthenticated]

    def _payload_from_plan(self, plan: PlanProduccion) -> list[tuple[Receta, Decimal]]:
        items_payload: list[tuple[Receta, Decimal]] = []
        for item in plan.items.select_related("receta").all():
            items_payload.append((item.receta, Decimal(str(item.cantidad or 0))))
        return items_payload

    def _payload_from_periodo(
        self,
        periodo: str,
        periodo_tipo: str,
    ) -> tuple[list[tuple[Receta, Decimal]], dict]:
        parsed = _parse_period(periodo)
        if not parsed:
            return [], {"periodo": periodo, "periodo_tipo": periodo_tipo, "planes_count": 0, "planes": []}

        year, month = parsed
        plans_qs = PlanProduccion.objects.filter(
            fecha_produccion__year=year,
            fecha_produccion__month=month,
        ).order_by("fecha_produccion", "id")
        if periodo_tipo == "q1":
            plans_qs = plans_qs.filter(fecha_produccion__day__lte=15)
        elif periodo_tipo == "q2":
            plans_qs = plans_qs.filter(fecha_produccion__day__gte=16)

        plans = list(plans_qs.only("id", "nombre", "fecha_produccion"))
        plan_ids = [p.id for p in plans]
        if not plan_ids:
            return [], {
                "periodo": f"{year:04d}-{month:02d}",
                "periodo_tipo": periodo_tipo,
                "planes_count": 0,
                "planes": [],
            }

        items_payload: list[tuple[Receta, Decimal]] = []
        items_qs = (
            PlanProduccionItem.objects.filter(plan_id__in=plan_ids)
            .select_related("receta")
            .only("receta_id", "cantidad", "plan_id", "receta__id", "receta__nombre")
        )
        for item in items_qs:
            items_payload.append((item.receta, Decimal(str(item.cantidad or 0))))

        return items_payload, {
            "periodo": f"{year:04d}-{month:02d}",
            "periodo_tipo": periodo_tipo,
            "planes_count": len(plans),
            "planes": [
                {
                    "id": p.id,
                    "nombre": p.nombre,
                    "fecha_produccion": str(p.fecha_produccion),
                }
                for p in plans[:50]
            ],
        }

    def _aggregate(self, items_payload: list[tuple[Receta, Decimal]]) -> dict:
        insumos = {}
        lineas_sin_match = 0
        lineas_sin_cantidad = 0
        lineas_sin_costo = 0

        for receta, factor in items_payload:
            if factor <= 0:
                continue
            for linea in receta.lineas.select_related("insumo", "insumo__unidad_base").all():
                if not linea.insumo_id:
                    lineas_sin_match += 1
                    continue

                qty = Decimal(str(linea.cantidad or 0))
                if qty <= 0:
                    lineas_sin_cantidad += 1
                    continue
                qty *= factor
                if qty <= 0:
                    continue

                unit_cost = Decimal(str(linea.costo_unitario_snapshot or 0))
                if unit_cost <= 0:
                    lineas_sin_costo += 1
                cost_total = qty * unit_cost if unit_cost > 0 else Decimal("0")

                row = insumos.setdefault(
                    linea.insumo_id,
                    {
                        "insumo_id": linea.insumo_id,
                        "insumo": linea.insumo.nombre,
                        "unidad": linea.unidad_texto or (linea.insumo.unidad_base.codigo if linea.insumo.unidad_base_id else ""),
                        "cantidad_requerida": Decimal("0"),
                        "costo_unitario": unit_cost,
                        "costo_total": Decimal("0"),
                    },
                )
                row["cantidad_requerida"] += qty
                row["costo_total"] += cost_total
                if row["costo_unitario"] <= 0 and unit_cost > 0:
                    row["costo_unitario"] = unit_cost

        existencias = {
            e.insumo_id: Decimal(str(e.stock_actual or 0))
            for e in ExistenciaInsumo.objects.filter(insumo_id__in=list(insumos.keys()))
        }

        alertas_capacidad = 0
        items = []
        for row in sorted(insumos.values(), key=lambda x: x["insumo"].lower()):
            stock = existencias.get(row["insumo_id"], Decimal("0"))
            faltante = row["cantidad_requerida"] - stock
            if faltante < 0:
                faltante = Decimal("0")
            alerta = faltante > 0
            if alerta:
                alertas_capacidad += 1
            items.append(
                {
                    "insumo_id": row["insumo_id"],
                    "insumo": row["insumo"],
                    "unidad": row["unidad"],
                    "cantidad_requerida": str(row["cantidad_requerida"]),
                    "stock_actual": str(stock),
                    "faltante": str(faltante),
                    "alerta_capacidad": alerta,
                    "costo_unitario": str(row["costo_unitario"]),
                    "costo_total": str(row["costo_total"]),
                }
            )

        return {
            "items": items,
            "totales": {
                "insumos": len(items),
                "costo_total": str(sum((Decimal(i["costo_total"]) for i in items), Decimal("0"))),
                "alertas_capacidad": alertas_capacidad,
                "lineas_sin_match": lineas_sin_match,
                "lineas_sin_cantidad": lineas_sin_cantidad,
                "lineas_sin_costo_unitario": lineas_sin_costo,
            },
        }

    def post(self, request):
        ser = MRPRequerimientosRequestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        plan_id = ser.validated_data.get("plan_id")
        periodo = (ser.validated_data.get("periodo") or "").strip()
        periodo_tipo = ser.validated_data.get("periodo_tipo") or "mes"
        raw_items = ser.validated_data.get("items") or []
        fecha_referencia = ser.validated_data.get("fecha_referencia")

        items_payload: list[tuple[Receta, Decimal]] = []
        source = "manual"
        plan = None
        periodo_scope = {
            "periodo": "",
            "periodo_tipo": periodo_tipo,
            "planes_count": 0,
            "planes": [],
        }

        if plan_id:
            plan = get_object_or_404(PlanProduccion, pk=plan_id)
            source = "plan"
            items_payload = self._payload_from_plan(plan)
            periodo_scope = {
                "periodo": plan.fecha_produccion.strftime("%Y-%m"),
                "periodo_tipo": "mes",
                "planes_count": 1,
                "planes": [
                    {
                        "id": plan.id,
                        "nombre": plan.nombre,
                        "fecha_produccion": str(plan.fecha_produccion),
                    }
                ],
            }
        elif periodo:
            source = "periodo"
            items_payload, periodo_scope = self._payload_from_periodo(periodo, periodo_tipo)
        else:
            for item in raw_items:
                receta = get_object_or_404(Receta, pk=item["receta_id"])
                items_payload.append((receta, Decimal(str(item["cantidad"]))))
            if fecha_referencia:
                periodo_scope["periodo"] = fecha_referencia.strftime("%Y-%m")

        data = self._aggregate(items_payload)
        response = {
            "source": source,
            "plan_id": plan.id if plan else None,
            "plan_nombre": plan.nombre if plan else "",
            "plan_fecha": str(plan.fecha_produccion) if plan else "",
            "periodo": periodo_scope["periodo"],
            "periodo_tipo": periodo_scope["periodo_tipo"],
            "planes_count": periodo_scope["planes_count"],
            "planes": periodo_scope["planes"],
            **data,
        }
        return Response(response, status=status.HTTP_200_OK)


class PlanProduccionListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _serialize_plan(plan: PlanProduccion, include_items: bool = False) -> dict:
        rows = list(plan.items.select_related("receta").all().order_by("id"))
        cantidad_total = sum((_to_decimal(r.cantidad) for r in rows), Decimal("0"))
        costo_total = sum((_to_decimal(r.costo_total_estimado) for r in rows), Decimal("0"))
        payload = {
            "id": plan.id,
            "nombre": plan.nombre,
            "fecha_produccion": str(plan.fecha_produccion),
            "notas": plan.notas or "",
            "items_count": len(rows),
            "cantidad_total": str(cantidad_total),
            "costo_total_estimado": str(costo_total.quantize(Decimal("0.001"))),
            "creado_por": plan.creado_por.username if plan.creado_por_id else "",
            "actualizado_en": plan.actualizado_en,
        }
        if include_items:
            payload["items"] = [
                {
                    "id": r.id,
                    "receta_id": r.receta_id,
                    "receta": r.receta.nombre,
                    "codigo_point": r.receta.codigo_point,
                    "cantidad": str(_to_decimal(r.cantidad)),
                    "costo_total_estimado": str(_to_decimal(r.costo_total_estimado).quantize(Decimal("0.001"))),
                    "notas": r.notas or "",
                }
                for r in rows
            ]
        return payload

    def get(self, request):
        if not request.user.has_perm("recetas.view_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para consultar planes de producción."},
                status=status.HTTP_403_FORBIDDEN,
            )

        q = (request.GET.get("q") or "").strip()
        periodo = (request.GET.get("periodo") or request.GET.get("mes") or "").strip()
        fecha_desde_raw = (request.GET.get("fecha_desde") or "").strip()
        fecha_hasta_raw = (request.GET.get("fecha_hasta") or "").strip()
        include_items = _parse_bool(request.GET.get("include_items"), default=False)
        limit = _parse_bounded_int(request.GET.get("limit", 120), default=120, min_value=1, max_value=400)

        fecha_desde = _parse_iso_date(fecha_desde_raw)
        if fecha_desde_raw and fecha_desde is None:
            return Response(
                {"detail": "fecha_desde inválida. Usa formato YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        fecha_hasta = _parse_iso_date(fecha_hasta_raw)
        if fecha_hasta_raw and fecha_hasta is None:
            return Response(
                {"detail": "fecha_hasta inválida. Usa formato YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if fecha_desde and fecha_hasta and fecha_hasta < fecha_desde:
            return Response(
                {"detail": "fecha_hasta no puede ser menor que fecha_desde."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = PlanProduccion.objects.prefetch_related("items__receta").order_by("-fecha_produccion", "-id")
        parsed_period = _parse_period(periodo)
        if periodo and not parsed_period:
            return Response(
                {"detail": "periodo inválido. Usa formato YYYY-MM."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if parsed_period:
            y, m = parsed_period
            qs = qs.filter(fecha_produccion__year=y, fecha_produccion__month=m)
            periodo = f"{y:04d}-{m:02d}"
        else:
            periodo = ""

        if fecha_desde:
            qs = qs.filter(fecha_produccion__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha_produccion__lte=fecha_hasta)
        if q:
            qs = qs.filter(
                Q(nombre__icontains=q) | Q(notas__icontains=q) | Q(items__receta__nombre__icontains=q)
            ).distinct()

        plans = list(qs[:limit])
        items = []
        items_count_total = 0
        cantidad_total = Decimal("0")
        costo_total = Decimal("0")
        for p in plans:
            serialized = self._serialize_plan(p, include_items=include_items)
            items.append(serialized)
            items_count_total += int(serialized["items_count"])
            cantidad_total += _to_decimal(serialized["cantidad_total"])
            costo_total += _to_decimal(serialized["costo_total_estimado"])

        return Response(
            {
                "filters": {
                    "q": q,
                    "periodo": periodo,
                    "fecha_desde": str(fecha_desde) if fecha_desde else "",
                    "fecha_hasta": str(fecha_hasta) if fecha_hasta else "",
                    "include_items": include_items,
                    "limit": limit,
                },
                "totales": {
                    "planes": len(items),
                    "renglones": items_count_total,
                    "cantidad_total": str(cantidad_total),
                    "costo_total_estimado": str(costo_total.quantize(Decimal("0.001"))),
                },
                "items": items,
            },
            status=status.HTTP_200_OK,
        )

    def post(self, request):
        if not request.user.has_perm("recetas.add_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para crear planes de producción."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = PlanProduccionCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        fecha_produccion = data.get("fecha_produccion") or timezone.localdate()
        nombre = (data.get("nombre") or "").strip()
        notas = (data.get("notas") or "").strip()
        rows = data["items"]

        receta_ids = sorted({int(row["receta_id"]) for row in rows})
        recetas = Receta.objects.filter(id__in=receta_ids).only("id", "nombre", "codigo_point")
        receta_map = {r.id: r for r in recetas}
        missing_ids = [rid for rid in receta_ids if rid not in receta_map]
        if missing_ids:
            return Response(
                {"detail": "Hay recetas inexistentes en items.", "missing_receta_ids": missing_ids},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            plan = PlanProduccion.objects.create(
                nombre=nombre or f"Plan {fecha_produccion} #{PlanProduccion.objects.count() + 1}",
                fecha_produccion=fecha_produccion,
                notas=notas,
                creado_por=request.user if request.user.is_authenticated else None,
            )
            for row in rows:
                receta = receta_map[int(row["receta_id"])]
                PlanProduccionItem.objects.create(
                    plan=plan,
                    receta=receta,
                    cantidad=_to_decimal(row.get("cantidad")),
                    notas=(row.get("notas") or "").strip()[:160],
                )

        payload = self._serialize_plan(plan, include_items=True)
        return Response(
            {"created": True, "plan": payload},
            status=status.HTTP_201_CREATED,
        )


class PlanProduccionDetailView(APIView):
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _serialize_item(row: PlanProduccionItem) -> dict:
        return {
            "id": row.id,
            "receta_id": row.receta_id,
            "receta": row.receta.nombre,
            "codigo_point": row.receta.codigo_point,
            "cantidad": str(_to_decimal(row.cantidad)),
            "costo_total_estimado": str(_to_decimal(row.costo_total_estimado).quantize(Decimal("0.001"))),
            "notas": row.notas or "",
        }

    @staticmethod
    def _plan_totals(plan: PlanProduccion) -> dict:
        rows = list(plan.items.select_related("receta").all())
        cantidad_total = sum((_to_decimal(r.cantidad) for r in rows), Decimal("0"))
        costo_total = sum((_to_decimal(r.costo_total_estimado) for r in rows), Decimal("0"))
        return {
            "items_count": len(rows),
            "cantidad_total": str(cantidad_total),
            "costo_total_estimado": str(costo_total.quantize(Decimal("0.001"))),
        }

    def get(self, request, plan_id: int):
        if not request.user.has_perm("recetas.view_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para consultar planes de producción."},
                status=status.HTTP_403_FORBIDDEN,
            )
        plan = get_object_or_404(PlanProduccion.objects.prefetch_related("items__receta"), pk=plan_id)
        totals = self._plan_totals(plan)
        items = [self._serialize_item(r) for r in plan.items.select_related("receta").all().order_by("id")]
        return Response(
            {
                "id": plan.id,
                "nombre": plan.nombre,
                "fecha_produccion": str(plan.fecha_produccion),
                "notas": plan.notas or "",
                "creado_por": plan.creado_por.username if plan.creado_por_id else "",
                "actualizado_en": plan.actualizado_en,
                "totals": totals,
                "items": items,
            },
            status=status.HTTP_200_OK,
        )

    def patch(self, request, plan_id: int):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para editar planes de producción."},
                status=status.HTTP_403_FORBIDDEN,
            )
        plan = get_object_or_404(PlanProduccion, pk=plan_id)
        ser = PlanProduccionUpdateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data
        update_fields = []
        if "nombre" in data:
            nombre = (data.get("nombre") or "").strip()
            if nombre:
                plan.nombre = nombre[:140]
                update_fields.append("nombre")
        if "fecha_produccion" in data:
            plan.fecha_produccion = data["fecha_produccion"]
            update_fields.append("fecha_produccion")
        if "notas" in data:
            plan.notas = (data.get("notas") or "").strip()
            update_fields.append("notas")
        if update_fields:
            plan.save(update_fields=update_fields + ["actualizado_en"])
        return Response(
            {
                "updated": bool(update_fields),
                "id": plan.id,
                "nombre": plan.nombre,
                "fecha_produccion": str(plan.fecha_produccion),
                "notas": plan.notas or "",
            },
            status=status.HTTP_200_OK,
        )

    def delete(self, request, plan_id: int):
        if not request.user.has_perm("recetas.delete_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para eliminar planes de producción."},
                status=status.HTTP_403_FORBIDDEN,
            )
        plan = get_object_or_404(PlanProduccion, pk=plan_id)
        payload = {"id": plan.id, "nombre": plan.nombre}
        plan.delete()
        return Response(
            {"deleted": True, "plan": payload},
            status=status.HTTP_200_OK,
        )


class PlanProduccionItemCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, plan_id: int):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para editar renglones de plan de producción."},
                status=status.HTTP_403_FORBIDDEN,
            )

        plan = get_object_or_404(PlanProduccion, pk=plan_id)
        ser = PlanProduccionItemCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data
        receta = Receta.objects.filter(pk=data["receta_id"]).only("id", "nombre", "codigo_point").first()
        if receta is None:
            return Response(
                {"detail": "receta_id no encontrada."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        item = PlanProduccionItem.objects.create(
            plan=plan,
            receta=receta,
            cantidad=_to_decimal(data["cantidad"]),
            notas=(data.get("notas") or "").strip()[:160],
        )
        return Response(
            {
                "created": True,
                "item": PlanProduccionDetailView._serialize_item(item),
                "plan_totals": PlanProduccionDetailView._plan_totals(plan),
            },
            status=status.HTTP_201_CREATED,
        )


class PlanProduccionItemDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, item_id: int):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para editar renglones de plan de producción."},
                status=status.HTTP_403_FORBIDDEN,
            )

        item = get_object_or_404(PlanProduccionItem.objects.select_related("plan", "receta"), pk=item_id)
        ser = PlanProduccionItemUpdateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data
        update_fields = []
        if "receta_id" in data:
            receta = Receta.objects.filter(pk=data["receta_id"]).only("id", "nombre", "codigo_point").first()
            if receta is None:
                return Response(
                    {"detail": "receta_id no encontrada."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            item.receta = receta
            update_fields.append("receta")
        if "cantidad" in data:
            item.cantidad = _to_decimal(data["cantidad"])
            update_fields.append("cantidad")
        if "notas" in data:
            item.notas = (data.get("notas") or "").strip()[:160]
            update_fields.append("notas")
        if update_fields:
            item.save(update_fields=update_fields)
        item.refresh_from_db()
        return Response(
            {
                "updated": bool(update_fields),
                "item": PlanProduccionDetailView._serialize_item(item),
                "plan_totals": PlanProduccionDetailView._plan_totals(item.plan),
            },
            status=status.HTTP_200_OK,
        )

    def delete(self, request, item_id: int):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para eliminar renglones de plan de producción."},
                status=status.HTTP_403_FORBIDDEN,
            )
        item = get_object_or_404(PlanProduccionItem.objects.select_related("plan"), pk=item_id)
        plan = item.plan
        payload = {"id": item.id, "plan_id": item.plan_id}
        item.delete()
        return Response(
            {
                "deleted": True,
                "item": payload,
                "plan_totals": PlanProduccionDetailView._plan_totals(plan),
            },
            status=status.HTTP_200_OK,
        )


class PlanDesdePronosticoCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.add_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para generar planes de producción."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = PlanDesdePronosticoRequestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        periodo = data["periodo"]
        incluir_preparaciones = bool(data.get("incluir_preparaciones"))
        nombre = (data.get("nombre") or "").strip() or f"Plan desde pronóstico {periodo}"
        fecha_produccion = data.get("fecha_produccion")
        if not fecha_produccion:
            fecha_produccion = date.fromisoformat(f"{periodo}-01")

        pronosticos_qs = (
            PronosticoVenta.objects.filter(periodo=periodo)
            .select_related("receta")
            .order_by("receta__nombre")
        )
        if not incluir_preparaciones:
            pronosticos_qs = pronosticos_qs.filter(receta__tipo=Receta.TIPO_PRODUCTO_FINAL)

        pronosticos = list(pronosticos_qs)
        if not pronosticos:
            return Response(
                {
                    "detail": "No hay pronósticos para generar plan en ese período con los filtros actuales.",
                    "periodo": periodo,
                    "incluir_preparaciones": incluir_preparaciones,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            plan = PlanProduccion.objects.create(
                nombre=nombre[:140],
                fecha_produccion=fecha_produccion,
                notas=f"Generado desde pronóstico {periodo}",
                creado_por=request.user if request.user.is_authenticated else None,
            )
            created = 0
            skipped = 0
            for p in pronosticos:
                qty = _to_decimal(getattr(p, "cantidad", 0))
                if qty <= 0:
                    skipped += 1
                    continue
                PlanProduccionItem.objects.create(
                    plan=plan,
                    receta=p.receta,
                    cantidad=qty,
                    notas=f"Pronóstico {periodo}",
                )
                created += 1

            if created == 0:
                plan.delete()
                return Response(
                    {
                        "detail": "No se creó plan: todos los pronósticos tenían cantidad 0.",
                        "periodo": periodo,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        return Response(
            {
                "plan_id": plan.id,
                "plan_nombre": plan.nombre,
                "plan_fecha": str(plan.fecha_produccion),
                "periodo": periodo,
                "incluir_preparaciones": incluir_preparaciones,
                "renglones_creados": created,
                "renglones_omitidos_cantidad_cero": skipped,
            },
            status=status.HTTP_201_CREATED,
        )


class ForecastBacktestView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.view_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para consultar backtest de pronóstico."},
                status=status.HTTP_403_FORBIDDEN,
            )
        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ser = ForecastBacktestRequestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        alcance = data.get("alcance") or "mes"
        fecha_base = data.get("fecha_base") or timezone.localdate()
        incluir_preparaciones = bool(data.get("incluir_preparaciones"))
        safety_pct = _to_decimal(data.get("safety_pct"), default=Decimal("0"))
        min_confianza_pct = _to_decimal(data.get("min_confianza_pct"), default=Decimal("0"))
        escenario = str(data.get("escenario") or "base").lower()
        periods = int(data.get("periods") or 3)
        top = int(data.get("top") or 10)

        sucursal = None
        sucursal_id = data.get("sucursal_id")
        if sucursal_id is not None:
            sucursal = Sucursal.objects.filter(pk=sucursal_id, activa=True).first()
            if sucursal is None:
                return Response(
                    {"detail": "Sucursal no encontrada o inactiva."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        windows = _forecast_backtest_windows(alcance, fecha_base, periods)
        if not windows:
            return Response(
                {"detail": "No se pudo construir ventanas de backtest."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        windows_payload: list[dict] = []
        sum_forecast_total = Decimal("0")
        sum_actual_total = Decimal("0")
        sum_abs_error = Decimal("0")
        ape_sum = Decimal("0")
        ape_count = 0

        for window_start, window_end in windows:
            periodo_window = f"{window_start.year:04d}-{window_start.month:02d}"
            forecast_result = _build_forecast_from_history(
                alcance=alcance,
                periodo=periodo_window,
                fecha_base=window_start,
                sucursal=sucursal,
                incluir_preparaciones=incluir_preparaciones,
                safety_pct=safety_pct,
            )
            forecast_result, _ = _filter_forecast_result_by_confianza(forecast_result, min_confianza_pct)
            qty_key = "forecast_qty"
            if escenario == "bajo":
                qty_key = "forecast_low"
            elif escenario == "alto":
                qty_key = "forecast_high"
            forecast_map = {
                int(row["receta_id"]): _to_decimal(row.get(qty_key))
                for row in (forecast_result.get("rows") or [])
            }

            actual_qs = VentaHistorica.objects.filter(fecha__gte=window_start, fecha__lte=window_end)
            if sucursal:
                actual_qs = actual_qs.filter(sucursal=sucursal)
            if not incluir_preparaciones:
                actual_qs = actual_qs.filter(receta__tipo=Receta.TIPO_PRODUCTO_FINAL)
            actual_map = {
                int(row["receta_id"]): _to_decimal(row["total"])
                for row in actual_qs.values("receta_id").annotate(total=Sum("cantidad"))
            }

            if min_confianza_pct > 0:
                union_ids = sorted(set(forecast_map.keys()))
            else:
                union_ids = sorted(set(forecast_map.keys()) | set(actual_map.keys()))
            if not union_ids:
                continue

            receta_names = {
                r.id: r.nombre for r in Receta.objects.filter(id__in=union_ids).only("id", "nombre")
            }
            rows = []
            forecast_total = Decimal("0")
            actual_total = Decimal("0")
            abs_error_total = Decimal("0")
            local_ape_sum = Decimal("0")
            local_ape_count = 0
            for receta_id in union_ids:
                forecast_qty = forecast_map.get(receta_id, Decimal("0"))
                actual_qty = actual_map.get(receta_id, Decimal("0"))
                delta_qty = forecast_qty - actual_qty
                abs_error = abs(delta_qty)

                forecast_total += forecast_qty
                actual_total += actual_qty
                abs_error_total += abs_error

                variacion_pct = None
                status_tag = "SIN_BASE"
                if actual_qty > 0:
                    variacion_pct = ((delta_qty / actual_qty) * Decimal("100")).quantize(Decimal("0.1"))
                    local_ape_sum += abs(variacion_pct)
                    local_ape_count += 1
                    if variacion_pct > Decimal("10"):
                        status_tag = "SOBRE"
                    elif variacion_pct < Decimal("-10"):
                        status_tag = "BAJO"
                    else:
                        status_tag = "OK"

                rows.append(
                    {
                        "receta_id": receta_id,
                        "receta": receta_names.get(receta_id) or f"Receta {receta_id}",
                        "forecast_qty": float(forecast_qty),
                        "actual_qty": float(actual_qty),
                        "delta_qty": float(delta_qty),
                        "abs_error": float(abs_error),
                        "variacion_pct": float(variacion_pct) if variacion_pct is not None else None,
                        "status": status_tag,
                    }
                )

            rows.sort(key=lambda r: abs(r["abs_error"]), reverse=True)
            mae = (abs_error_total / Decimal(str(len(union_ids)))).quantize(Decimal("0.001"))
            mape = None
            if local_ape_count > 0:
                mape = (local_ape_sum / Decimal(str(local_ape_count))).quantize(Decimal("0.1"))
                ape_sum += local_ape_sum
                ape_count += local_ape_count

            sum_forecast_total += forecast_total
            sum_actual_total += actual_total
            sum_abs_error += abs_error_total

            windows_payload.append(
                {
                    "window_start": str(window_start),
                    "window_end": str(window_end),
                    "periodo": periodo_window,
                    "recetas_count": len(union_ids),
                    "forecast_total": float(forecast_total),
                    "actual_total": float(actual_total),
                    "bias_total": float((forecast_total - actual_total).quantize(Decimal("0.001"))),
                    "mae": float(mae),
                    "mape": float(mape) if mape is not None else None,
                    "top_errors": rows[:top],
                }
            )

        if not windows_payload:
            return Response(
                {"detail": "No hay historial suficiente para evaluar backtest en el alcance seleccionado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        overall_mape = None
        if ape_count > 0:
            overall_mape = (ape_sum / Decimal(str(ape_count))).quantize(Decimal("0.1"))
        overall_mae = (sum_abs_error / Decimal(str(max(1, len(windows_payload))))).quantize(Decimal("0.001"))

        payload = {
            "scope": {
                "alcance": alcance,
                "fecha_base": str(fecha_base),
                "periods": periods,
                "escenario": escenario,
                "min_confianza_pct": _to_float(min_confianza_pct),
                "sucursal_id": sucursal.id if sucursal else None,
                "sucursal_nombre": f"{sucursal.codigo} - {sucursal.nombre}" if sucursal else "Todas",
            },
            "totals": {
                "windows_evaluated": len(windows_payload),
                "forecast_total": float(sum_forecast_total),
                "actual_total": float(sum_actual_total),
                "bias_total": float((sum_forecast_total - sum_actual_total).quantize(Decimal("0.001"))),
                "mae_promedio": float(overall_mae),
                "mape_promedio": float(overall_mape) if overall_mape is not None else None,
            },
            "windows": windows_payload,
        }
        if export_format:
            return _forecast_backtest_export_response(payload, export_format)
        return Response(payload, status=status.HTTP_200_OK)


class ForecastInsightsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.has_perm("recetas.view_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para consultar insights de pronóstico."},
                status=status.HTTP_403_FORBIDDEN,
            )

        months = _parse_bounded_int(request.GET.get("months", 12), default=12, min_value=1, max_value=36)
        top = _parse_bounded_int(request.GET.get("top", 20), default=20, min_value=1, max_value=100)
        offset_top = _parse_bounded_int(request.GET.get("offset_top", 0), default=0, min_value=0, max_value=5000)
        incluir_preparaciones = _parse_bool(request.GET.get("incluir_preparaciones"), default=False)
        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "Parámetro export inválido. Usa 'csv' o 'xlsx'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        fecha_hasta_raw = (request.GET.get("fecha_hasta") or "").strip()
        fecha_hasta = _parse_iso_date(fecha_hasta_raw)
        if fecha_hasta_raw and fecha_hasta is None:
            return Response(
                {"detail": "fecha_hasta inválida. Usa formato YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if fecha_hasta is None:
            fecha_hasta = timezone.localdate()
        fecha_desde = fecha_hasta - timedelta(days=(months * 31) - 1)

        sucursal = None
        sucursal_id_raw = (request.GET.get("sucursal_id") or "").strip()
        if sucursal_id_raw:
            if not sucursal_id_raw.isdigit():
                return Response(
                    {"detail": "sucursal_id debe ser numérico."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            sucursal = Sucursal.objects.filter(pk=int(sucursal_id_raw), activa=True).first()
            if sucursal is None:
                return Response(
                    {"detail": "Sucursal no encontrada o inactiva."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        receta = None
        receta_id_raw = (request.GET.get("receta_id") or "").strip()
        if receta_id_raw:
            if not receta_id_raw.isdigit():
                return Response(
                    {"detail": "receta_id debe ser numérico."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            receta = Receta.objects.filter(pk=int(receta_id_raw)).first()
            if receta is None:
                return Response(
                    {"detail": "Receta no encontrada."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        qs = VentaHistorica.objects.filter(fecha__gte=fecha_desde, fecha__lte=fecha_hasta)
        if sucursal:
            qs = qs.filter(sucursal=sucursal)
        if receta:
            qs = qs.filter(receta=receta)
        if not incluir_preparaciones:
            qs = qs.filter(receta__tipo=Receta.TIPO_PRODUCTO_FINAL)

        rows = list(
            qs.values("fecha", "receta_id", "receta__nombre")
            .annotate(total=Sum("cantidad"))
            .order_by("fecha", "receta_id")
        )
        if not rows:
            payload = {
                "scope": {
                    "months": months,
                    "fecha_desde": str(fecha_desde),
                    "fecha_hasta": str(fecha_hasta),
                    "sucursal_id": sucursal.id if sucursal else None,
                    "sucursal": sucursal.nombre if sucursal else "Todas",
                    "receta_id": receta.id if receta else None,
                    "receta": receta.nombre if receta else "Todas",
                    "top": top,
                    "offset_top": offset_top,
                },
                "totales": {
                    "filas": 0,
                    "dias_con_venta": 0,
                    "recetas": 0,
                    "top_recetas_total": 0,
                    "top_recetas_returned": 0,
                    "cantidad_total": 0.0,
                    "promedio_diario": 0.0,
                },
                "seasonality": {"by_month": [], "by_weekday": []},
                "top_recetas": [],
            }
            if export_format:
                return _forecast_insights_export_response(payload, export_format)
            return Response(payload, status=status.HTTP_200_OK)

        date_totals: dict[date, Decimal] = defaultdict(lambda: Decimal("0"))
        recipe_totals: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
        recipe_days: dict[int, set[date]] = defaultdict(set)
        recipe_names: dict[int, str] = {}

        for row in rows:
            d = row["fecha"]
            receta_id = int(row["receta_id"])
            qty = _to_decimal(row["total"])
            date_totals[d] += qty
            recipe_totals[receta_id] += qty
            recipe_days[receta_id].add(d)
            recipe_names[receta_id] = row["receta__nombre"]

        daily_values = list(date_totals.values())
        total_qty = sum(daily_values, Decimal("0"))
        global_avg = total_qty / Decimal(str(len(daily_values) or 1))

        month_map: dict[int, list[Decimal]] = defaultdict(list)
        weekday_map: dict[int, list[Decimal]] = defaultdict(list)
        for d, qty in date_totals.items():
            month_map[d.month].append(qty)
            weekday_map[d.weekday()].append(qty)

        month_labels = {
            1: "Ene",
            2: "Feb",
            3: "Mar",
            4: "Abr",
            5: "May",
            6: "Jun",
            7: "Jul",
            8: "Ago",
            9: "Sep",
            10: "Oct",
            11: "Nov",
            12: "Dic",
        }
        weekday_labels = ["Lun", "Mar", "Mie", "Jue", "Vie", "Sab", "Dom"]

        month_rows = []
        for month in sorted(month_map.keys()):
            samples = month_map[month]
            avg_qty = sum(samples, Decimal("0")) / Decimal(str(len(samples) or 1))
            index_pct = Decimal("100")
            if global_avg > 0:
                index_pct = (avg_qty / global_avg) * Decimal("100")
            month_rows.append(
                {
                    "month": month,
                    "label": month_labels.get(month, str(month)),
                    "samples": len(samples),
                    "avg_qty": _to_float(avg_qty),
                    "index_pct": _to_float(index_pct.quantize(Decimal("0.1"))),
                }
            )

        weekday_rows = []
        for wd in range(7):
            samples = weekday_map.get(wd, [])
            avg_qty = Decimal("0")
            index_pct = Decimal("0")
            if samples:
                avg_qty = sum(samples, Decimal("0")) / Decimal(str(len(samples)))
                index_pct = Decimal("100")
                if global_avg > 0:
                    index_pct = (avg_qty / global_avg) * Decimal("100")
            weekday_rows.append(
                {
                    "weekday": wd,
                    "label": weekday_labels[wd],
                    "samples": len(samples),
                    "avg_qty": _to_float(avg_qty),
                    "index_pct": _to_float(index_pct.quantize(Decimal("0.1"))) if samples else 0.0,
                }
            )

        top_recetas_all = []
        for receta_id, qty_total in sorted(recipe_totals.items(), key=lambda item: item[1], reverse=True):
            days_count = len(recipe_days.get(receta_id, set())) or 1
            avg_day = qty_total / Decimal(str(days_count))
            share = Decimal("0")
            if total_qty > 0:
                share = (qty_total / total_qty) * Decimal("100")
            top_recetas_all.append(
                {
                    "receta_id": receta_id,
                    "receta": recipe_names.get(receta_id) or f"Receta {receta_id}",
                    "cantidad_total": _to_float(qty_total),
                    "promedio_dia_activo": _to_float(avg_day.quantize(Decimal("0.001"))),
                    "dias_con_venta": days_count,
                    "participacion_pct": _to_float(share.quantize(Decimal("0.1"))),
                }
            )
        top_recetas = top_recetas_all[offset_top : offset_top + top]

        payload = {
            "scope": {
                "months": months,
                "fecha_desde": str(fecha_desde),
                "fecha_hasta": str(fecha_hasta),
                "sucursal_id": sucursal.id if sucursal else None,
                "sucursal": sucursal.nombre if sucursal else "Todas",
                "receta_id": receta.id if receta else None,
                "receta": receta.nombre if receta else "Todas",
                "top": top,
                "offset_top": offset_top,
            },
            "totales": {
                "filas": len(rows),
                "dias_con_venta": len(date_totals),
                "recetas": len(recipe_totals),
                "top_recetas_total": len(top_recetas_all),
                "top_recetas_returned": len(top_recetas),
                "cantidad_total": _to_float(total_qty),
                "promedio_diario": _to_float(global_avg.quantize(Decimal("0.001"))),
            },
            "seasonality": {
                "by_month": month_rows,
                "by_weekday": weekday_rows,
            },
            "top_recetas": top_recetas,
        }
        if export_format:
            return _forecast_insights_export_response(payload, export_format)
        return Response(payload, status=status.HTTP_200_OK)


class VentaHistoricaListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.has_perm("recetas.view_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para consultar historial de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )
        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        q = (request.GET.get("q") or "").strip()
        periodo = (request.GET.get("periodo") or request.GET.get("mes") or "").strip()
        sucursal_id_raw = (request.GET.get("sucursal_id") or "").strip()
        receta_id_raw = (request.GET.get("receta_id") or "").strip()
        fecha_desde_raw = (request.GET.get("fecha_desde") or "").strip()
        fecha_hasta_raw = (request.GET.get("fecha_hasta") or "").strip()
        limit = _parse_bounded_int(request.GET.get("limit", 150), default=150, min_value=1, max_value=1000)
        offset = _parse_bounded_int(request.GET.get("offset", 0), default=0, min_value=0, max_value=200000)

        fecha_desde = _parse_iso_date(fecha_desde_raw)
        if fecha_desde_raw and fecha_desde is None:
            return Response(
                {"detail": "fecha_desde inválida. Usa formato YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        fecha_hasta = _parse_iso_date(fecha_hasta_raw)
        if fecha_hasta_raw and fecha_hasta is None:
            return Response(
                {"detail": "fecha_hasta inválida. Usa formato YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if fecha_desde and fecha_hasta and fecha_hasta < fecha_desde:
            return Response(
                {"detail": "fecha_hasta no puede ser menor que fecha_desde."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = VentaHistorica.objects.select_related("receta", "sucursal").order_by("-fecha", "-id")
        parsed_period = _parse_period(periodo)
        if periodo and not parsed_period:
            return Response(
                {"detail": "periodo inválido. Usa formato YYYY-MM."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if parsed_period:
            year, month = parsed_period
            qs = qs.filter(fecha__year=year, fecha__month=month)
            periodo = f"{year:04d}-{month:02d}"
        else:
            periodo = ""

        if fecha_desde:
            qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha__lte=fecha_hasta)

        if sucursal_id_raw:
            if not sucursal_id_raw.isdigit():
                return Response(
                    {"detail": "sucursal_id debe ser numérico."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            qs = qs.filter(sucursal_id=int(sucursal_id_raw))

        if receta_id_raw:
            if not receta_id_raw.isdigit():
                return Response(
                    {"detail": "receta_id debe ser numérico."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            qs = qs.filter(receta_id=int(receta_id_raw))

        if q:
            qs = qs.filter(
                Q(receta__nombre__icontains=q)
                | Q(receta__codigo_point__icontains=q)
                | Q(sucursal__nombre__icontains=q)
                | Q(sucursal__codigo__icontains=q)
                | Q(fuente__icontains=q)
            )

        total_rows = qs.count()
        agg = qs.aggregate(
            cantidad_total=Sum("cantidad"),
            tickets_total=Sum("tickets"),
            monto_total=Sum("monto_total"),
        )
        cantidad_total = _to_decimal(agg.get("cantidad_total"))
        tickets_total = int(agg.get("tickets_total") or 0)
        monto_total = _to_decimal(agg.get("monto_total"))
        by_sucursal: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
        for row in qs.values("sucursal__codigo").annotate(total=Sum("cantidad")):
            sucursal_key = str(row.get("sucursal__codigo") or "GLOBAL")
            by_sucursal[sucursal_key] += _to_decimal(row.get("total"))

        rows = list(qs[offset : offset + limit])
        items = []

        for r in rows:
            cantidad = _to_decimal(r.cantidad)
            monto = _to_decimal(r.monto_total)
            items.append(
                {
                    "id": r.id,
                    "fecha": str(r.fecha),
                    "receta_id": r.receta_id,
                    "receta": r.receta.nombre,
                    "codigo_point": r.receta.codigo_point,
                    "sucursal_id": r.sucursal_id,
                    "sucursal": r.sucursal.nombre if r.sucursal_id and r.sucursal else "",
                    "sucursal_codigo": r.sucursal.codigo if r.sucursal_id and r.sucursal else "",
                    "cantidad": str(cantidad),
                    "tickets": int(r.tickets or 0),
                    "monto_total": str(monto),
                    "fuente": r.fuente,
                    "actualizado_en": r.actualizado_en,
                }
            )

        sucursales_payload = [
            {"sucursal": k, "cantidad_total": str(v)}
            for k, v in sorted(by_sucursal.items(), key=lambda entry: entry[1], reverse=True)
        ]

        payload = {
            "filters": {
                "q": q,
                "periodo": periodo,
                "sucursal_id": sucursal_id_raw,
                "receta_id": receta_id_raw,
                "fecha_desde": str(fecha_desde) if fecha_desde else "",
                "fecha_hasta": str(fecha_hasta) if fecha_hasta else "",
                "limit": limit,
                "offset": offset,
            },
            "totales": {
                "rows": len(items),
                "rows_total": int(total_rows),
                "rows_returned": len(items),
                "cantidad_total": str(cantidad_total),
                "tickets_total": tickets_total,
                "monto_total": str(monto_total),
                "by_sucursal": sucursales_payload,
            },
            "items": items,
        }
        if export_format:
            return _ventas_historial_export_response(payload, export_format)
        return Response(payload, status=status.HTTP_200_OK)


class PronosticoVentaListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.has_perm("recetas.view_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para consultar pronósticos de venta."},
                status=status.HTTP_403_FORBIDDEN,
            )
        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        q = (request.GET.get("q") or "").strip()
        periodo = (request.GET.get("periodo") or "").strip()
        periodo_desde = (request.GET.get("periodo_desde") or "").strip()
        periodo_hasta = (request.GET.get("periodo_hasta") or "").strip()
        receta_id_raw = (request.GET.get("receta_id") or "").strip()
        limit = _parse_bounded_int(request.GET.get("limit", 150), default=150, min_value=1, max_value=1000)
        offset = _parse_bounded_int(request.GET.get("offset", 0), default=0, min_value=0, max_value=200000)

        if periodo:
            parsed_period = _parse_period(periodo)
            if not parsed_period:
                return Response(
                    {"detail": "periodo inválido. Usa formato YYYY-MM."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            year, month = parsed_period
            periodo = f"{year:04d}-{month:02d}"

        if periodo_desde:
            parsed_since = _parse_period(periodo_desde)
            if not parsed_since:
                return Response(
                    {"detail": "periodo_desde inválido. Usa formato YYYY-MM."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            periodo_desde = f"{parsed_since[0]:04d}-{parsed_since[1]:02d}"

        if periodo_hasta:
            parsed_until = _parse_period(periodo_hasta)
            if not parsed_until:
                return Response(
                    {"detail": "periodo_hasta inválido. Usa formato YYYY-MM."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            periodo_hasta = f"{parsed_until[0]:04d}-{parsed_until[1]:02d}"

        if periodo_desde and periodo_hasta and periodo_hasta < periodo_desde:
            return Response(
                {"detail": "periodo_hasta no puede ser menor que periodo_desde."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = PronosticoVenta.objects.select_related("receta").order_by("-periodo", "receta__nombre")
        if periodo:
            qs = qs.filter(periodo=periodo)
        if periodo_desde:
            qs = qs.filter(periodo__gte=periodo_desde)
        if periodo_hasta:
            qs = qs.filter(periodo__lte=periodo_hasta)

        if receta_id_raw:
            if not receta_id_raw.isdigit():
                return Response(
                    {"detail": "receta_id debe ser numérico."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            qs = qs.filter(receta_id=int(receta_id_raw))

        if q:
            qs = qs.filter(
                Q(receta__nombre__icontains=q)
                | Q(receta__codigo_point__icontains=q)
                | Q(fuente__icontains=q)
            )

        total_rows = qs.count()
        cantidad_total = _to_decimal(qs.aggregate(total=Sum("cantidad")).get("total"))
        periodos_count = qs.values("periodo").distinct().count()
        rows = list(qs[offset : offset + limit])
        items = []
        for r in rows:
            qty = _to_decimal(r.cantidad)
            items.append(
                {
                    "id": r.id,
                    "receta_id": r.receta_id,
                    "receta": r.receta.nombre,
                    "codigo_point": r.receta.codigo_point,
                    "periodo": r.periodo,
                    "cantidad": str(qty),
                    "fuente": r.fuente,
                    "actualizado_en": r.actualizado_en,
                }
            )

        payload = {
            "filters": {
                "q": q,
                "periodo": periodo,
                "periodo_desde": periodo_desde,
                "periodo_hasta": periodo_hasta,
                "receta_id": receta_id_raw,
                "limit": limit,
                "offset": offset,
            },
            "totales": {
                "rows": len(items),
                "rows_total": int(total_rows),
                "rows_returned": len(items),
                "cantidad_total": str(cantidad_total),
                "periodos_count": int(periodos_count),
            },
            "items": items,
        }
        if export_format:
            return _ventas_pronostico_export_response(payload, export_format)
        return Response(payload, status=status.HTTP_200_OK)


class VentasPipelineResumenView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.has_perm("recetas.view_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para consultar resumen del pipeline de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )

        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        periodo = (request.GET.get("periodo") or "").strip()
        top = _parse_bounded_int(request.GET.get("top", 120), default=120, min_value=1, max_value=500)
        offset = _parse_bounded_int(request.GET.get("offset", 0), default=0, min_value=0, max_value=5000)
        top_sucursales = _parse_bounded_int(request.GET.get("top_sucursales", 120), default=120, min_value=1, max_value=500)
        offset_sucursales = _parse_bounded_int(
            request.GET.get("offset_sucursales", 0),
            default=0,
            min_value=0,
            max_value=5000,
        )
        status_filter = (request.GET.get("status") or "").strip().upper()
        allowed_status = {"", "SOBRE", "BAJO", "OK", "SIN_SOLICITUD", "SIN_MOV", "DESVIADAS"}
        if status_filter not in allowed_status:
            return Response(
                {"detail": "status inválido. Usa SOBRE, BAJO, OK, SIN_SOLICITUD, SIN_MOV o DESVIADAS."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        delta_min = _to_decimal(request.GET.get("delta_min"), default=Decimal("0"))
        if delta_min < 0:
            delta_min = Decimal("0")
        sort_by = (request.GET.get("sort_by") or "delta_abs").strip().lower()
        sort_dir = (request.GET.get("sort_dir") or "desc").strip().lower()
        sort_sucursales_by = (request.GET.get("sort_sucursales_by") or "delta_abs").strip().lower()
        sort_sucursales_dir = (request.GET.get("sort_sucursales_dir") or "desc").strip().lower()
        q = (request.GET.get("q") or "").strip()

        allowed_sort_rows = {
            "delta_abs",
            "delta",
            "historial",
            "solicitud",
            "pronostico",
            "cobertura_pct",
            "cumplimiento_pct",
            "receta",
        }
        allowed_sort_sucursales = {"delta_abs", "delta", "historial", "solicitud", "cumplimiento_pct", "sucursal"}
        if sort_by not in allowed_sort_rows:
            return Response(
                {
                    "detail": (
                        "sort_by inválido. Usa delta_abs, delta, historial, solicitud, pronostico, cobertura_pct, "
                        "cumplimiento_pct o receta."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        if sort_sucursales_by not in allowed_sort_sucursales:
            return Response(
                {"detail": "sort_sucursales_by inválido. Usa delta_abs, delta, historial, solicitud, cumplimiento_pct o sucursal."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if sort_dir not in {"asc", "desc"}:
            return Response(
                {"detail": "sort_dir inválido. Usa asc o desc."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if sort_sucursales_dir not in {"asc", "desc"}:
            return Response(
                {"detail": "sort_sucursales_dir inválido. Usa asc o desc."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        def _status_match(tag: str) -> bool:
            if not status_filter:
                return True
            if status_filter == "DESVIADAS":
                return tag in {"SOBRE", "BAJO"}
            return tag == status_filter

        if periodo:
            parsed_period = _parse_period(periodo)
            if not parsed_period:
                return Response(
                    {"detail": "periodo inválido. Usa formato YYYY-MM."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            periodo = f"{parsed_period[0]:04d}-{parsed_period[1]:02d}"
        else:
            today = timezone.localdate()
            periodo = f"{today.year:04d}-{today.month:02d}"
            parsed_period = (today.year, today.month)
        year, month = parsed_period

        incluir_preparaciones = _parse_bool(request.GET.get("incluir_preparaciones"), default=False)
        sucursal = None
        sucursal_id_raw = (request.GET.get("sucursal_id") or "").strip()
        if sucursal_id_raw:
            if not sucursal_id_raw.isdigit():
                return Response(
                    {"detail": "sucursal_id debe ser numérico."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            sucursal = Sucursal.objects.filter(pk=int(sucursal_id_raw), activa=True).first()
            if sucursal is None:
                return Response(
                    {"detail": "Sucursal no encontrada o inactiva."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        historial_qs = VentaHistorica.objects.filter(fecha__year=year, fecha__month=month)
        solicitudes_qs = SolicitudVenta.objects.filter(periodo=periodo)
        pronostico_qs = PronosticoVenta.objects.filter(periodo=periodo)

        if q:
            historial_qs = historial_qs.filter(
                Q(receta__nombre__icontains=q)
                | Q(receta__codigo_point__icontains=q)
            )
            solicitudes_qs = solicitudes_qs.filter(
                Q(receta__nombre__icontains=q)
                | Q(receta__codigo_point__icontains=q)
            )
            pronostico_qs = pronostico_qs.filter(
                Q(receta__nombre__icontains=q)
                | Q(receta__codigo_point__icontains=q)
            )

        if sucursal:
            historial_qs = historial_qs.filter(sucursal=sucursal)
            solicitudes_qs = solicitudes_qs.filter(sucursal=sucursal)

        if not incluir_preparaciones:
            historial_qs = historial_qs.filter(receta__tipo=Receta.TIPO_PRODUCTO_FINAL)
            solicitudes_qs = solicitudes_qs.filter(receta__tipo=Receta.TIPO_PRODUCTO_FINAL)
            pronostico_qs = pronostico_qs.filter(receta__tipo=Receta.TIPO_PRODUCTO_FINAL)

        historial_total = _to_decimal(historial_qs.aggregate(total=Sum("cantidad"))["total"])
        solicitud_total = _to_decimal(solicitudes_qs.aggregate(total=Sum("cantidad"))["total"])
        pronostico_total = _to_decimal(pronostico_qs.aggregate(total=Sum("cantidad"))["total"])

        by_alcance = {k: Decimal("0") for k in [SolicitudVenta.ALCANCE_MES, SolicitudVenta.ALCANCE_SEMANA, SolicitudVenta.ALCANCE_FIN_SEMANA]}
        for row in solicitudes_qs.values("alcance").annotate(total=Sum("cantidad")):
            by_alcance[row["alcance"]] = _to_decimal(row["total"])

        historial_sucursal_map = {
            int(row["sucursal_id"]): _to_decimal(row["total"])
            for row in historial_qs.values("sucursal_id").annotate(total=Sum("cantidad"))
            if row.get("sucursal_id") is not None
        }
        solicitud_sucursal_map = {
            int(row["sucursal_id"]): _to_decimal(row["total"])
            for row in solicitudes_qs.values("sucursal_id").annotate(total=Sum("cantidad"))
            if row.get("sucursal_id") is not None
        }
        sucursal_ids = sorted(set(historial_sucursal_map.keys()) | set(solicitud_sucursal_map.keys()))
        sucursal_map = {
            int(sid): (codigo, nombre)
            for sid, codigo, nombre in Sucursal.objects.filter(id__in=sucursal_ids).values_list("id", "codigo", "nombre")
        }
        by_sucursal_rows_tmp: list[dict[str, Any]] = []
        for sid in sucursal_ids:
            hist_qty = historial_sucursal_map.get(sid, Decimal("0"))
            sol_qty = solicitud_sucursal_map.get(sid, Decimal("0"))
            delta = hist_qty - sol_qty
            cumplimiento_pct = None
            if sol_qty > 0:
                cumplimiento_pct = ((hist_qty / sol_qty) * Decimal("100")).quantize(Decimal("0.1"))
            if sol_qty > 0:
                threshold = max(Decimal("1"), sol_qty * Decimal("0.10"))
                if delta > threshold:
                    status_tag = "SOBRE"
                elif delta < (threshold * Decimal("-1")):
                    status_tag = "BAJO"
                else:
                    status_tag = "OK"
            elif hist_qty > 0:
                status_tag = "SIN_SOLICITUD"
            else:
                status_tag = "SIN_MOV"

            suc_codigo, suc_nombre = sucursal_map.get(sid, ("", f"Sucursal {sid}"))
            delta_abs = abs(delta)
            by_sucursal_rows_tmp.append(
                {
                    "_delta_abs": _to_float(delta_abs),
                    "sucursal_id": sid,
                    "sucursal_codigo": suc_codigo,
                    "sucursal": suc_nombre,
                    "historial_qty": _to_float(hist_qty),
                    "solicitud_qty": _to_float(sol_qty),
                    "delta_historial_vs_solicitud": _to_float(delta.quantize(Decimal("0.001"))),
                    "cumplimiento_pct": _to_float(cumplimiento_pct) if cumplimiento_pct is not None else None,
                    "status": status_tag,
                }
            )
        by_sucursal_rows_tmp = [
            row
            for row in by_sucursal_rows_tmp
            if _status_match(str(row.get("status") or "")) and Decimal(str(row["_delta_abs"])) >= delta_min
        ]
        by_sucursal_sort_field = {
            "delta_abs": "_delta_abs",
            "delta": "delta_historial_vs_solicitud",
            "historial": "historial_qty",
            "solicitud": "solicitud_qty",
            "cumplimiento_pct": "cumplimiento_pct",
            "sucursal": "sucursal",
        }[sort_sucursales_by]
        for row in by_sucursal_rows_tmp:
            sort_value = row.get(by_sucursal_sort_field)
            if isinstance(sort_value, str):
                row["_sort"] = sort_value.lower()
            elif sort_value is None:
                row["_sort"] = float("-inf")
            else:
                row["_sort"] = _to_float(sort_value)
        by_sucursal_rows_tmp.sort(key=lambda row: row["_sort"], reverse=(sort_sucursales_dir == "desc"))
        by_sucursal_filtered = len(by_sucursal_rows_tmp)
        by_sucursal_rows = []
        for row in by_sucursal_rows_tmp[offset_sucursales : offset_sucursales + top_sucursales]:
            row.pop("_sort", None)
            row.pop("_delta_abs", None)
            by_sucursal_rows.append(row)
        by_sucursal_status_counts = {
            "SOBRE": sum(1 for row in by_sucursal_rows_tmp if row.get("status") == "SOBRE"),
            "BAJO": sum(1 for row in by_sucursal_rows_tmp if row.get("status") == "BAJO"),
            "OK": sum(1 for row in by_sucursal_rows_tmp if row.get("status") == "OK"),
            "SIN_SOLICITUD": sum(1 for row in by_sucursal_rows_tmp if row.get("status") == "SIN_SOLICITUD"),
            "SIN_MOV": sum(1 for row in by_sucursal_rows_tmp if row.get("status") == "SIN_MOV"),
        }

        cobertura_solicitud_pct = None
        if pronostico_total > 0:
            cobertura_solicitud_pct = ((solicitud_total / pronostico_total) * Decimal("100")).quantize(Decimal("0.1"))

        cumplimiento_historial_pct = None
        if solicitud_total > 0:
            cumplimiento_historial_pct = ((historial_total / solicitud_total) * Decimal("100")).quantize(Decimal("0.1"))

        latest_historial = historial_qs.order_by("-actualizado_en").values_list("actualizado_en", flat=True).first()
        latest_pronostico = pronostico_qs.order_by("-actualizado_en").values_list("actualizado_en", flat=True).first()
        latest_solicitud = solicitudes_qs.order_by("-actualizado_en").values_list("actualizado_en", flat=True).first()

        historial_map = {
            int(row["receta_id"]): _to_decimal(row["total"])
            for row in historial_qs.values("receta_id").annotate(total=Sum("cantidad"))
        }
        pronostico_map = {
            int(row["receta_id"]): _to_decimal(row["total"])
            for row in pronostico_qs.values("receta_id").annotate(total=Sum("cantidad"))
        }
        solicitud_map = {
            int(row["receta_id"]): _to_decimal(row["total"])
            for row in solicitudes_qs.values("receta_id").annotate(total=Sum("cantidad"))
        }
        receta_ids = sorted(set(historial_map.keys()) | set(pronostico_map.keys()) | set(solicitud_map.keys()))
        receta_names = {
            int(receta_id): nombre
            for receta_id, nombre in Receta.objects.filter(id__in=receta_ids).values_list("id", "nombre")
        }

        rows_tmp: list[dict[str, Any]] = []
        for receta_id in receta_ids:
            historial_qty = historial_map.get(receta_id, Decimal("0"))
            pronostico_qty = pronostico_map.get(receta_id, Decimal("0"))
            solicitud_qty = solicitud_map.get(receta_id, Decimal("0"))
            delta_solicitud_vs_pronostico = solicitud_qty - pronostico_qty
            delta_historial_vs_solicitud = historial_qty - solicitud_qty

            cobertura_pct = None
            if pronostico_qty > 0:
                cobertura_pct = ((solicitud_qty / pronostico_qty) * Decimal("100")).quantize(Decimal("0.1"))

            cumplimiento_pct = None
            if solicitud_qty > 0:
                cumplimiento_pct = ((historial_qty / solicitud_qty) * Decimal("100")).quantize(Decimal("0.1"))

            status_tag = "SIN_MOV"
            if solicitud_qty > 0:
                threshold = max(Decimal("1"), solicitud_qty * Decimal("0.10"))
                if delta_historial_vs_solicitud > threshold:
                    status_tag = "SOBRE"
                elif delta_historial_vs_solicitud < (threshold * Decimal("-1")):
                    status_tag = "BAJO"
                else:
                    status_tag = "OK"
            elif historial_qty > 0:
                status_tag = "SIN_SOLICITUD"
            elif pronostico_qty > 0:
                status_tag = "SIN_SOLICITUD"

            rows_tmp.append(
                {
                    "_delta_abs": _to_float(abs(delta_historial_vs_solicitud)),
                    "receta_id": receta_id,
                    "receta": receta_names.get(receta_id) or f"Receta {receta_id}",
                    "historial_qty": _to_float(historial_qty),
                    "pronostico_qty": _to_float(pronostico_qty),
                    "solicitud_qty": _to_float(solicitud_qty),
                    "delta_solicitud_vs_pronostico": _to_float(delta_solicitud_vs_pronostico.quantize(Decimal("0.001"))),
                    "delta_historial_vs_solicitud": _to_float(delta_historial_vs_solicitud.quantize(Decimal("0.001"))),
                    "cobertura_pct": _to_float(cobertura_pct) if cobertura_pct is not None else None,
                    "cumplimiento_pct": _to_float(cumplimiento_pct) if cumplimiento_pct is not None else None,
                    "status": status_tag,
                }
            )
        rows_tmp = [
            row
            for row in rows_tmp
            if _status_match(str(row.get("status") or "")) and Decimal(str(row["_delta_abs"])) >= delta_min
        ]
        row_sort_field = {
            "delta_abs": "_delta_abs",
            "delta": "delta_historial_vs_solicitud",
            "historial": "historial_qty",
            "solicitud": "solicitud_qty",
            "pronostico": "pronostico_qty",
            "cobertura_pct": "cobertura_pct",
            "cumplimiento_pct": "cumplimiento_pct",
            "receta": "receta",
        }[sort_by]
        for row in rows_tmp:
            sort_value = row.get(row_sort_field)
            if isinstance(sort_value, str):
                row["_sort"] = sort_value.lower()
            elif sort_value is None:
                row["_sort"] = float("-inf")
            else:
                row["_sort"] = _to_float(sort_value)
        rows_tmp.sort(key=lambda row: row["_sort"], reverse=(sort_dir == "desc"))
        rows_filtered = len(rows_tmp)
        rows = []
        for row in rows_tmp[offset : offset + top]:
            row.pop("_sort", None)
            row.pop("_delta_abs", None)
            rows.append(row)
        rows_status_counts = {
            "SOBRE": sum(1 for row in rows_tmp if row.get("status") == "SOBRE"),
            "BAJO": sum(1 for row in rows_tmp if row.get("status") == "BAJO"),
            "OK": sum(1 for row in rows_tmp if row.get("status") == "OK"),
            "SIN_SOLICITUD": sum(1 for row in rows_tmp if row.get("status") == "SIN_SOLICITUD"),
            "SIN_MOV": sum(1 for row in rows_tmp if row.get("status") == "SIN_MOV"),
        }

        payload = {
            "scope": {
                "periodo": periodo,
                "sucursal_id": sucursal.id if sucursal else None,
                "sucursal": sucursal.nombre if sucursal else "Todas",
                "incluir_preparaciones": incluir_preparaciones,
                "top": top,
                "offset": offset,
                "top_sucursales": top_sucursales,
                "offset_sucursales": offset_sucursales,
                "sort_by": sort_by,
                "sort_dir": sort_dir,
                "sort_sucursales_by": sort_sucursales_by,
                "sort_sucursales_dir": sort_sucursales_dir,
                "status": status_filter or "",
                "delta_min": _to_float(delta_min),
                "q": q,
            },
            "totales": {
                "historial_qty": _to_float(historial_total),
                "pronostico_qty": _to_float(pronostico_total),
                "solicitud_qty": _to_float(solicitud_total),
                "delta_solicitud_vs_pronostico": _to_float((solicitud_total - pronostico_total).quantize(Decimal("0.001"))),
                "delta_historial_vs_solicitud": _to_float((historial_total - solicitud_total).quantize(Decimal("0.001"))),
                "cobertura_solicitud_pct": _to_float(cobertura_solicitud_pct) if cobertura_solicitud_pct is not None else None,
                "cumplimiento_historial_pct": _to_float(cumplimiento_historial_pct) if cumplimiento_historial_pct is not None else None,
                "historial_recetas": historial_qs.values("receta_id").distinct().count(),
                "pronostico_recetas": pronostico_qs.values("receta_id").distinct().count(),
                "solicitud_recetas": solicitudes_qs.values("receta_id").distinct().count(),
                "rows_count": len(receta_ids),
                "rows_filtered": rows_filtered,
                "rows_returned": len(rows),
                "by_sucursal_filtered": by_sucursal_filtered,
                "by_sucursal_returned": len(by_sucursal_rows),
                "rows_status": rows_status_counts,
                "by_sucursal_status": by_sucursal_status_counts,
            },
            "solicitud_by_alcance": {
                "MES": _to_float(by_alcance[SolicitudVenta.ALCANCE_MES]),
                "SEMANA": _to_float(by_alcance[SolicitudVenta.ALCANCE_SEMANA]),
                "FIN_SEMANA": _to_float(by_alcance[SolicitudVenta.ALCANCE_FIN_SEMANA]),
            },
            "by_sucursal": by_sucursal_rows,
            "latest_updates": {
                "historial": latest_historial,
                "pronostico": latest_pronostico,
                "solicitud": latest_solicitud,
            },
            "rows": rows,
        }
        if export_format:
            return _ventas_pipeline_export_response(payload, export_format)
        return Response(payload, status=status.HTTP_200_OK)


class SolicitudVentaListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.has_perm("recetas.view_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para consultar solicitudes de venta."},
                status=status.HTTP_403_FORBIDDEN,
            )
        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        q = (request.GET.get("q") or "").strip()
        periodo = (request.GET.get("periodo") or "").strip()
        alcance = (request.GET.get("alcance") or "").strip().upper()
        sucursal_id_raw = (request.GET.get("sucursal_id") or "").strip()
        receta_id_raw = (request.GET.get("receta_id") or "").strip()
        fecha_desde_raw = (request.GET.get("fecha_desde") or "").strip()
        fecha_hasta_raw = (request.GET.get("fecha_hasta") or "").strip()
        limit = _parse_bounded_int(request.GET.get("limit", 150), default=150, min_value=1, max_value=1000)
        offset = _parse_bounded_int(request.GET.get("offset", 0), default=0, min_value=0, max_value=200000)
        include_forecast_ref = _parse_bool(request.GET.get("include_forecast_ref"), default=False)
        forecast_status_filter = (request.GET.get("forecast_status") or "").strip().upper()
        forecast_delta_min = _to_decimal(request.GET.get("forecast_delta_min"), default=Decimal("0"))
        sort_by = (request.GET.get("sort_by") or "fecha_inicio").strip().lower()
        sort_dir = (request.GET.get("sort_dir") or "desc").strip().lower()
        if forecast_delta_min < 0:
            forecast_delta_min = Decimal("0")
        allowed_forecast_status = {"", "SOBRE", "BAJO", "OK", "SIN_FORECAST", "DESVIADAS"}
        if forecast_status_filter not in allowed_forecast_status:
            return Response(
                {"detail": "forecast_status inválido. Usa SOBRE, BAJO, OK, SIN_FORECAST o DESVIADAS."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        allowed_sort_by = {
            "fecha_inicio",
            "fecha_fin",
            "cantidad",
            "receta",
            "sucursal",
            "alcance",
            "periodo",
            "forecast_delta",
            "forecast_status",
        }
        if sort_by not in allowed_sort_by:
            return Response(
                {
                    "detail": (
                        "sort_by inválido. Usa fecha_inicio, fecha_fin, cantidad, receta, sucursal, alcance, periodo, "
                        "forecast_delta o forecast_status."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        if sort_dir not in {"asc", "desc"}:
            return Response(
                {"detail": "sort_dir inválido. Usa asc o desc."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if forecast_status_filter or forecast_delta_min > 0:
            include_forecast_ref = True
        if sort_by in {"forecast_delta", "forecast_status"}:
            include_forecast_ref = True

        if periodo:
            parsed_period = _parse_period(periodo)
            if not parsed_period:
                return Response(
                    {"detail": "periodo inválido. Usa formato YYYY-MM."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            periodo = f"{parsed_period[0]:04d}-{parsed_period[1]:02d}"

        allowed_alcance = {
            SolicitudVenta.ALCANCE_MES,
            SolicitudVenta.ALCANCE_SEMANA,
            SolicitudVenta.ALCANCE_FIN_SEMANA,
        }
        if alcance and alcance not in allowed_alcance:
            return Response(
                {"detail": "alcance inválido. Usa MES, SEMANA o FIN_SEMANA."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        fecha_desde = _parse_iso_date(fecha_desde_raw)
        if fecha_desde_raw and fecha_desde is None:
            return Response(
                {"detail": "fecha_desde inválida. Usa formato YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        fecha_hasta = _parse_iso_date(fecha_hasta_raw)
        if fecha_hasta_raw and fecha_hasta is None:
            return Response(
                {"detail": "fecha_hasta inválida. Usa formato YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if fecha_desde and fecha_hasta and fecha_hasta < fecha_desde:
            return Response(
                {"detail": "fecha_hasta no puede ser menor que fecha_desde."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = SolicitudVenta.objects.select_related("receta", "sucursal").order_by("-fecha_inicio", "receta__nombre")
        if periodo:
            qs = qs.filter(periodo=periodo)
        if alcance:
            qs = qs.filter(alcance=alcance)
        if fecha_desde:
            qs = qs.filter(fecha_inicio__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha_fin__lte=fecha_hasta)

        if sucursal_id_raw:
            if not sucursal_id_raw.isdigit():
                return Response(
                    {"detail": "sucursal_id debe ser numérico."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            qs = qs.filter(sucursal_id=int(sucursal_id_raw))

        if receta_id_raw:
            if not receta_id_raw.isdigit():
                return Response(
                    {"detail": "receta_id debe ser numérico."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            qs = qs.filter(receta_id=int(receta_id_raw))

        if q:
            qs = qs.filter(
                Q(receta__nombre__icontains=q)
                | Q(receta__codigo_point__icontains=q)
                | Q(sucursal__nombre__icontains=q)
                | Q(sucursal__codigo__icontains=q)
                | Q(fuente__icontains=q)
            )

        rows = list(qs)
        pronostico_map: dict[tuple[int, str], Decimal] = {}
        if include_forecast_ref and rows:
            receta_ids = {int(r.receta_id) for r in rows}
            periodos = {str(r.periodo) for r in rows if r.periodo}
            if receta_ids and periodos:
                for p in (
                    PronosticoVenta.objects.filter(receta_id__in=receta_ids, periodo__in=periodos)
                    .values("receta_id", "periodo")
                    .annotate(total=Sum("cantidad"))
                ):
                    pronostico_map[(int(p["receta_id"]), str(p["periodo"]))] = _to_decimal(p["total"])

        items = []
        cantidad_total = Decimal("0")
        by_alcance = {k: 0 for k in allowed_alcance}
        forecast_status_counts = {"SOBRE": 0, "BAJO": 0, "OK": 0, "SIN_FORECAST": 0}

        def _forecast_match(status_value: str, delta_value: Decimal) -> bool:
            if forecast_status_filter:
                if forecast_status_filter == "DESVIADAS":
                    if status_value not in {"SOBRE", "BAJO"}:
                        return False
                elif status_value != forecast_status_filter:
                    return False
            if forecast_delta_min > 0 and abs(delta_value) < forecast_delta_min:
                return False
            return True

        for r in rows:
            qty = _to_decimal(r.cantidad)
            item = {
                "id": r.id,
                "receta_id": r.receta_id,
                "receta": r.receta.nombre,
                "codigo_point": r.receta.codigo_point,
                "sucursal_id": r.sucursal_id,
                "sucursal": r.sucursal.nombre if r.sucursal_id and r.sucursal else "",
                "sucursal_codigo": r.sucursal.codigo if r.sucursal_id and r.sucursal else "",
                "alcance": r.alcance,
                "periodo": r.periodo,
                "fecha_inicio": str(r.fecha_inicio),
                "fecha_fin": str(r.fecha_fin),
                "cantidad": str(qty),
                "fuente": r.fuente,
                "actualizado_en": r.actualizado_en,
            }
            if include_forecast_ref:
                forecast_qty = _to_decimal(pronostico_map.get((int(r.receta_id), str(r.periodo))), default=Decimal("0"))
                if forecast_qty > 0:
                    delta_forecast = qty - forecast_qty
                    threshold = max(Decimal("1"), forecast_qty * Decimal("0.10"))
                    if delta_forecast > threshold:
                        forecast_status = "SOBRE"
                    elif delta_forecast < (threshold * Decimal("-1")):
                        forecast_status = "BAJO"
                    else:
                        forecast_status = "OK"
                else:
                    delta_forecast = qty
                    forecast_status = "SIN_FORECAST"

                item["forecast_ref"] = {
                    "status": forecast_status,
                    "forecast_qty": _to_float(forecast_qty),
                    "solicitud_qty": _to_float(qty),
                    "delta_solicitud_vs_forecast": _to_float(delta_forecast.quantize(Decimal("0.001"))),
                }
                if not _forecast_match(forecast_status, delta_forecast):
                    continue
                forecast_status_counts[forecast_status] = forecast_status_counts.get(forecast_status, 0) + 1
            cantidad_total += qty
            by_alcance[r.alcance] = by_alcance.get(r.alcance, 0) + 1
            items.append(item)

        def _sort_value(item: dict[str, Any]) -> Any:
            if sort_by == "cantidad":
                return _to_float(_to_decimal(item.get("cantidad"), default=Decimal("0")))
            if sort_by == "fecha_inicio":
                return _parse_iso_date(str(item.get("fecha_inicio") or "")) or date.min
            if sort_by == "fecha_fin":
                return _parse_iso_date(str(item.get("fecha_fin") or "")) or date.min
            if sort_by == "receta":
                return str(item.get("receta") or "").lower()
            if sort_by == "sucursal":
                return str(item.get("sucursal") or "").lower()
            if sort_by == "alcance":
                return str(item.get("alcance") or "")
            if sort_by == "periodo":
                return str(item.get("periodo") or "")
            if sort_by == "forecast_delta":
                return _to_float((item.get("forecast_ref") or {}).get("delta_solicitud_vs_forecast") or 0)
            if sort_by == "forecast_status":
                status_rank = {"SOBRE": 4, "BAJO": 3, "OK": 2, "SIN_FORECAST": 1}
                return int(status_rank.get(str((item.get("forecast_ref") or {}).get("status") or ""), 0))
            return str(item.get("fecha_inicio") or "")

        items.sort(key=_sort_value, reverse=(sort_dir == "desc"))
        rows_total = len(items)
        items = items[offset : offset + limit]

        payload = {
            "filters": {
                "q": q,
                "periodo": periodo,
                "alcance": alcance,
                "sucursal_id": sucursal_id_raw,
                "receta_id": receta_id_raw,
                "fecha_desde": str(fecha_desde) if fecha_desde else "",
                "fecha_hasta": str(fecha_hasta) if fecha_hasta else "",
                "limit": limit,
                "offset": offset,
                "include_forecast_ref": include_forecast_ref,
                "forecast_status": forecast_status_filter or "",
                "forecast_delta_min": _to_float(forecast_delta_min),
                "sort_by": sort_by,
                "sort_dir": sort_dir,
            },
            "totales": {
                "rows": len(items),
                "rows_total": rows_total,
                "rows_returned": len(items),
                "cantidad_total": str(cantidad_total),
                "by_alcance": by_alcance,
                "forecast_ref_status": forecast_status_counts if include_forecast_ref else {},
            },
            "items": items,
        }
        if export_format:
            return _ventas_solicitudes_export_response(payload, export_format)
        return Response(payload, status=status.HTTP_200_OK)


def _process_pronostico_venta_bulk(
    data: dict[str, Any],
    *,
    dry_run_override: bool | None = None,
) -> dict[str, Any]:
    rows = data["rows"]
    modo = data.get("modo") or "replace"
    fuente = (data.get("fuente") or "API_PRON_BULK").strip()[:40] or "API_PRON_BULK"
    dry_run = bool(data.get("dry_run", True) if dry_run_override is None else dry_run_override)
    stop_on_error = bool(data.get("stop_on_error", False))
    top = int(data.get("top") or 120)

    receta_cache: dict[tuple[int, str, str], Receta | None] = {}
    created = 0
    updated = 0
    skipped = 0
    terminated_early = False
    result_rows: list[dict] = []

    tx_cm = nullcontext() if dry_run else transaction.atomic()
    with tx_cm:
        for index, row in enumerate(rows, start=1):
            receta_id = row.get("receta_id")
            receta_name = str(row.get("receta") or "").strip()
            codigo_point = str(row.get("codigo_point") or "").strip()
            receta = _resolve_receta_bulk_ref(
                receta_id=receta_id,
                receta_name=receta_name,
                codigo_point=codigo_point,
                cache=receta_cache,
            )
            if receta is None:
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "receta_not_found",
                        "receta_id": int(receta_id or 0) or None,
                        "receta_input": receta_name or codigo_point,
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            periodo = _normalize_periodo_mes(row.get("periodo"))
            cantidad = _to_decimal(row.get("cantidad"), default=Decimal("0"))
            existing = PronosticoVenta.objects.filter(receta=receta, periodo=periodo).first()

            previous_qty = _to_decimal(existing.cantidad, default=Decimal("0")) if existing else Decimal("0")
            if existing:
                new_qty = previous_qty + cantidad if modo == "accumulate" else cantidad
                action = "UPDATED"
                updated += 1
            else:
                new_qty = cantidad
                action = "CREATED"
                created += 1

            if not dry_run:
                if existing:
                    existing.cantidad = new_qty
                    existing.fuente = fuente
                    existing.save(update_fields=["cantidad", "fuente", "actualizado_en"])
                else:
                    PronosticoVenta.objects.create(
                        receta=receta,
                        periodo=periodo,
                        cantidad=new_qty,
                        fuente=fuente,
                    )

            result_rows.append(
                {
                    "row": index,
                    "status": action,
                    "receta_id": receta.id,
                    "receta": receta.nombre,
                    "periodo": periodo,
                    "cantidad_prev": float(previous_qty),
                    "cantidad_nueva": float(new_qty),
                }
            )

    error_count = sum(1 for row in result_rows if row.get("status") == "ERROR")
    return {
        "dry_run": dry_run,
        "mode": modo,
        "fuente": fuente,
        "terminated_early": terminated_early,
        "summary": {
            "total_rows": len(rows),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": error_count,
            "applied": 0 if dry_run else (created + updated),
        },
        "rows": result_rows[:top],
    }


class PronosticoVentaImportPreviewView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_pronosticoventa"):
            return Response(
                {"detail": "No tienes permisos para previsualizar pronósticos de venta."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = PronosticoVentaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        payload = _process_pronostico_venta_bulk(ser.validated_data, dry_run_override=True)
        payload["preview"] = True
        return Response(payload, status=status.HTTP_200_OK)


class PronosticoVentaImportConfirmView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_pronosticoventa"):
            return Response(
                {"detail": "No tienes permisos para confirmar importación de pronósticos de venta."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = PronosticoVentaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        payload = _process_pronostico_venta_bulk(ser.validated_data, dry_run_override=False)
        payload["preview"] = False
        return Response(payload, status=status.HTTP_200_OK)


class PronosticoVentaBulkUpsertView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_pronosticoventa"):
            return Response(
                {"detail": "No tienes permisos para importar pronósticos de venta."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = PronosticoVentaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        payload = _process_pronostico_venta_bulk(ser.validated_data)
        return Response(payload, status=status.HTTP_200_OK)


def _process_venta_historica_bulk(
    data: dict[str, Any],
    *,
    dry_run_override: bool | None = None,
) -> dict[str, Any]:
    rows = data["rows"]
    modo = data.get("modo") or "replace"
    fuente = (data.get("fuente") or "API_VENTAS_BULK").strip()[:40] or "API_VENTAS_BULK"
    dry_run = bool(data.get("dry_run", True) if dry_run_override is None else dry_run_override)
    stop_on_error = bool(data.get("stop_on_error", False))
    top = int(data.get("top") or 120)

    default_sucursal = None
    default_sucursal_id = data.get("sucursal_default_id")
    if default_sucursal_id is not None:
        default_sucursal = Sucursal.objects.filter(pk=default_sucursal_id, activa=True).first()
        if default_sucursal is None:
            raise ValueError("Sucursal default no encontrada o inactiva.")

    receta_cache: dict[tuple[int, str, str], Receta | None] = {}
    sucursal_cache: dict[tuple[int, str, str, int], Sucursal | None] = {}
    created = 0
    updated = 0
    skipped = 0
    terminated_early = False
    result_rows: list[dict] = []

    tx_cm = nullcontext() if dry_run else transaction.atomic()
    with tx_cm:
        for index, row in enumerate(rows, start=1):
            receta_id = row.get("receta_id")
            receta_name = str(row.get("receta") or "").strip()
            codigo_point = str(row.get("codigo_point") or "").strip()
            receta = _resolve_receta_bulk_ref(
                receta_id=receta_id,
                receta_name=receta_name,
                codigo_point=codigo_point,
                cache=receta_cache,
            )
            if receta is None:
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "receta_not_found",
                        "receta_id": int(receta_id or 0) or None,
                        "receta_input": receta_name or codigo_point,
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            sucursal_id = row.get("sucursal_id")
            sucursal_name = str(row.get("sucursal") or "").strip()
            sucursal_codigo = str(row.get("sucursal_codigo") or "").strip()
            has_sucursal_ref = bool(sucursal_id) or bool(sucursal_name) or bool(sucursal_codigo)
            sucursal = _resolve_sucursal_bulk_ref(
                sucursal_id=sucursal_id,
                sucursal_name=sucursal_name,
                sucursal_codigo=sucursal_codigo,
                default_sucursal=default_sucursal,
                cache=sucursal_cache,
            )
            if has_sucursal_ref and sucursal is None:
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "sucursal_not_found",
                        "receta_id": receta.id,
                        "receta": receta.nombre,
                        "sucursal_input": sucursal_codigo or sucursal_name or str(sucursal_id),
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            fecha = row["fecha"]
            cantidad = _to_decimal(row.get("cantidad"), default=Decimal("0"))
            tickets = int(row.get("tickets") or 0)
            monto_total_raw = row.get("monto_total", None)
            monto_total = _to_decimal(monto_total_raw) if monto_total_raw is not None else None
            if monto_total is not None and monto_total <= 0:
                monto_total = None

            existing_qs = VentaHistorica.objects.filter(receta=receta, fecha=fecha)
            if sucursal:
                existing_qs = existing_qs.filter(sucursal=sucursal)
            else:
                existing_qs = existing_qs.filter(sucursal__isnull=True)
            existing = existing_qs.order_by("id").first()

            previous_qty = _to_decimal(existing.cantidad, default=Decimal("0")) if existing else Decimal("0")
            previous_tickets = int(existing.tickets or 0) if existing else 0
            if existing:
                if modo == "accumulate":
                    new_qty = previous_qty + cantidad
                    new_tickets = previous_tickets + tickets
                else:
                    new_qty = cantidad
                    new_tickets = tickets
                action = "UPDATED"
                updated += 1
            else:
                new_qty = cantidad
                new_tickets = tickets
                action = "CREATED"
                created += 1

            if not dry_run:
                if existing:
                    existing.cantidad = new_qty
                    existing.tickets = new_tickets
                    if modo == "accumulate":
                        if monto_total is not None:
                            existing.monto_total = _to_decimal(existing.monto_total, default=Decimal("0")) + monto_total
                    else:
                        existing.monto_total = monto_total
                    existing.fuente = fuente
                    existing.save(update_fields=["cantidad", "tickets", "monto_total", "fuente", "actualizado_en"])
                else:
                    VentaHistorica.objects.create(
                        receta=receta,
                        sucursal=sucursal,
                        fecha=fecha,
                        cantidad=new_qty,
                        tickets=new_tickets,
                        monto_total=monto_total,
                        fuente=fuente,
                    )

            result_rows.append(
                {
                    "row": index,
                    "status": action,
                    "receta_id": receta.id,
                    "receta": receta.nombre,
                    "sucursal_id": sucursal.id if sucursal else None,
                    "sucursal": sucursal.nombre if sucursal else "",
                    "fecha": str(fecha),
                    "cantidad_prev": float(previous_qty),
                    "cantidad_nueva": float(new_qty),
                    "tickets_prev": previous_tickets,
                    "tickets_nuevos": new_tickets,
                }
            )

    error_count = sum(1 for row in result_rows if row.get("status") == "ERROR")
    return {
        "dry_run": dry_run,
        "mode": modo,
        "fuente": fuente,
        "terminated_early": terminated_early,
        "summary": {
            "total_rows": len(rows),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": error_count,
            "applied": 0 if dry_run else (created + updated),
        },
        "rows": result_rows[:top],
    }


class VentaHistoricaImportPreviewView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para previsualizar historial de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = VentaHistoricaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_venta_historica_bulk(ser.validated_data, dry_run_override=True)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        payload["preview"] = True
        return Response(payload, status=status.HTTP_200_OK)


class VentaHistoricaImportConfirmView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para confirmar importación de historial de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = VentaHistoricaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_venta_historica_bulk(ser.validated_data, dry_run_override=False)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        payload["preview"] = False
        return Response(payload, status=status.HTTP_200_OK)


class VentaHistoricaBulkUpsertView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para importar historial de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = VentaHistoricaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_venta_historica_bulk(ser.validated_data)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(payload, status=status.HTTP_200_OK)


def _process_solicitud_venta_bulk(
    data: dict[str, Any],
    *,
    dry_run_override: bool | None = None,
) -> dict[str, Any]:
    rows = data["rows"]
    modo = data.get("modo") or "replace"
    fuente = (data.get("fuente") or "API_SOL_BULK").strip()[:40] or "API_SOL_BULK"
    dry_run = bool(data.get("dry_run", True) if dry_run_override is None else dry_run_override)
    stop_on_error = bool(data.get("stop_on_error", False))
    top = int(data.get("top") or 120)

    default_sucursal = None
    default_sucursal_id = data.get("sucursal_default_id")
    if default_sucursal_id is not None:
        default_sucursal = Sucursal.objects.filter(pk=default_sucursal_id, activa=True).first()
        if default_sucursal is None:
            raise ValueError("Sucursal default no encontrada o inactiva.")

    receta_cache: dict[tuple[int, str, str], Receta | None] = {}
    sucursal_cache: dict[tuple[int, str, str, int], Sucursal | None] = {}
    created = 0
    updated = 0
    skipped = 0
    terminated_early = False
    result_rows: list[dict] = []

    tx_cm = nullcontext() if dry_run else transaction.atomic()
    with tx_cm:
        for index, row in enumerate(rows, start=1):
            receta_id = row.get("receta_id")
            receta_name = str(row.get("receta") or "").strip()
            codigo_point = str(row.get("codigo_point") or "").strip()
            receta = _resolve_receta_bulk_ref(
                receta_id=receta_id,
                receta_name=receta_name,
                codigo_point=codigo_point,
                cache=receta_cache,
            )
            if receta is None:
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "receta_not_found",
                        "receta_id": int(receta_id or 0) or None,
                        "receta_input": receta_name or codigo_point,
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            sucursal_id = row.get("sucursal_id")
            sucursal_name = str(row.get("sucursal") or "").strip()
            sucursal_codigo = str(row.get("sucursal_codigo") or "").strip()
            has_sucursal_ref = bool(sucursal_id) or bool(sucursal_name) or bool(sucursal_codigo)
            sucursal = _resolve_sucursal_bulk_ref(
                sucursal_id=sucursal_id,
                sucursal_name=sucursal_name,
                sucursal_codigo=sucursal_codigo,
                default_sucursal=default_sucursal,
                cache=sucursal_cache,
            )
            if has_sucursal_ref and sucursal is None:
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "sucursal_not_found",
                        "receta_id": receta.id,
                        "receta": receta.nombre,
                        "sucursal_input": sucursal_codigo or sucursal_name or str(sucursal_id),
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            alcance_ui = str(row.get("alcance") or "mes").strip().lower()
            alcance_model = _ui_to_model_alcance(alcance_ui)
            periodo_default = _normalize_periodo_mes(row.get("periodo"))
            fecha_base_default = row.get("fecha_base") or timezone.localdate()
            periodo, fecha_inicio, fecha_fin = _resolve_solicitud_window(
                alcance=alcance_model,
                periodo_raw=row.get("periodo"),
                fecha_base_raw=row.get("fecha_base"),
                fecha_inicio_raw=row.get("fecha_inicio"),
                fecha_fin_raw=row.get("fecha_fin"),
                periodo_default=periodo_default,
                fecha_base_default=fecha_base_default,
            )
            cantidad = _to_decimal(row.get("cantidad"), default=Decimal("0"))
            if cantidad <= 0:
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "invalid_qty",
                        "receta_id": receta.id,
                        "receta": receta.nombre,
                        "cantidad": float(cantidad),
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            existing = SolicitudVenta.objects.filter(
                receta=receta,
                sucursal=sucursal,
                alcance=alcance_model,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
            ).first()

            previous_qty = _to_decimal(existing.cantidad, default=Decimal("0")) if existing else Decimal("0")
            if existing:
                new_qty = previous_qty + cantidad if modo == "accumulate" else cantidad
                action = "UPDATED"
                updated += 1
            else:
                new_qty = cantidad
                action = "CREATED"
                created += 1

            if not dry_run:
                if existing:
                    existing.periodo = periodo
                    existing.cantidad = new_qty
                    existing.fuente = fuente
                    existing.save(update_fields=["periodo", "cantidad", "fuente", "actualizado_en"])
                else:
                    SolicitudVenta.objects.create(
                        receta=receta,
                        sucursal=sucursal,
                        alcance=alcance_model,
                        periodo=periodo,
                        fecha_inicio=fecha_inicio,
                        fecha_fin=fecha_fin,
                        cantidad=new_qty,
                        fuente=fuente,
                    )

            result_rows.append(
                {
                    "row": index,
                    "status": action,
                    "receta_id": receta.id,
                    "receta": receta.nombre,
                    "sucursal_id": sucursal.id if sucursal else None,
                    "sucursal": sucursal.nombre if sucursal else "",
                    "alcance": alcance_model,
                    "periodo": periodo,
                    "fecha_inicio": str(fecha_inicio),
                    "fecha_fin": str(fecha_fin),
                    "cantidad_prev": float(previous_qty),
                    "cantidad_nueva": float(new_qty),
                }
            )

    error_count = sum(1 for row in result_rows if row.get("status") == "ERROR")
    return {
        "dry_run": dry_run,
        "mode": modo,
        "fuente": fuente,
        "terminated_early": terminated_early,
        "summary": {
            "total_rows": len(rows),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": error_count,
            "applied": 0 if dry_run else (created + updated),
        },
        "rows": result_rows[:top],
    }


class SolicitudVentaImportPreviewView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para previsualizar solicitudes de venta."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = SolicitudVentaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_solicitud_venta_bulk(ser.validated_data, dry_run_override=True)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        payload["preview"] = True
        return Response(payload, status=status.HTTP_200_OK)


class SolicitudVentaImportConfirmView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para confirmar importación de solicitudes de venta."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = SolicitudVentaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_solicitud_venta_bulk(ser.validated_data, dry_run_override=False)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        payload["preview"] = False
        return Response(payload, status=status.HTTP_200_OK)


class SolicitudVentaBulkUpsertView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para importar solicitudes de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = SolicitudVentaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_solicitud_venta_bulk(ser.validated_data)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(payload, status=status.HTTP_200_OK)


class ForecastEstadisticoView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.view_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para consultar pronóstico estadístico."},
                status=status.HTTP_403_FORBIDDEN,
            )
        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ser = ForecastEstadisticoRequestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        periodo = _normalize_periodo_mes(data.get("periodo"))
        fecha_base = data.get("fecha_base") or timezone.localdate()
        alcance = data.get("alcance") or "mes"
        incluir_preparaciones = bool(data.get("incluir_preparaciones"))
        safety_pct = _to_decimal(data.get("safety_pct"), default=Decimal("0"))
        min_confianza_pct = _to_decimal(data.get("min_confianza_pct"), default=Decimal("0"))
        escenario_compare = str(data.get("escenario_compare") or "base").lower()
        top = int(data.get("top") or 120)

        sucursal = None
        sucursal_id = data.get("sucursal_id")
        if sucursal_id is not None:
            sucursal = Sucursal.objects.filter(pk=sucursal_id, activa=True).first()
            if sucursal is None:
                return Response(
                    {"detail": "Sucursal no encontrada o inactiva."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        result = _build_forecast_from_history(
            alcance=alcance,
            periodo=periodo,
            fecha_base=fecha_base,
            sucursal=sucursal,
            incluir_preparaciones=incluir_preparaciones,
            safety_pct=safety_pct,
        )
        result, filtered_conf = _filter_forecast_result_by_confianza(result, min_confianza_pct)
        result["min_confianza_pct"] = min_confianza_pct
        if not result.get("rows"):
            return Response(
                {"detail": "No hay forecast tras aplicar el filtro de confianza mínima."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        payload = _forecast_session_payload(result, top_rows=top)

        compare_payload = None
        if bool(data.get("include_solicitud_compare", True)):
            full_payload = _forecast_session_payload(result, top_rows=max(len(result.get("rows") or []), 1))
            compare_payload = _serialize_forecast_compare(
                _forecast_vs_solicitud_preview(full_payload, escenario=escenario_compare),
                top=top,
            )

        result_payload = {
            "scope": {
                "alcance": payload["alcance"],
                "periodo": payload["periodo"],
                "target_start": payload["target_start"],
                "target_end": payload["target_end"],
                "sucursal_nombre": payload["sucursal_nombre"],
                "sucursal_id": payload.get("sucursal_id"),
                "escenario_compare": escenario_compare,
                "min_confianza_pct": _to_float(min_confianza_pct),
                "filtered_conf": filtered_conf,
            },
            "totals": payload["totals"],
            "rows": payload["rows"],
            "compare_solicitud": compare_payload,
        }
        if export_format:
            return _forecast_estadistico_export_response(result_payload, export_format)
        return Response(result_payload, status=status.HTTP_200_OK)


class ForecastEstadisticoGuardarView(APIView):
    permission_classes = [IsAuthenticated]

    _ESCENARIO_TO_KEY = {
        "base": "forecast_qty",
        "bajo": "forecast_low",
        "alto": "forecast_high",
    }

    def post(self, request):
        if not request.user.has_perm("recetas.change_pronosticoventa"):
            return Response(
                {"detail": "No tienes permisos para guardar pronóstico estadístico."},
                status=status.HTTP_403_FORBIDDEN,
            )
        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ser = ForecastEstadisticoGuardarSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        periodo = _normalize_periodo_mes(data.get("periodo"))
        fecha_base = data.get("fecha_base") or timezone.localdate()
        alcance = data.get("alcance") or "mes"
        incluir_preparaciones = bool(data.get("incluir_preparaciones"))
        safety_pct = _to_decimal(data.get("safety_pct"), default=Decimal("0"))
        min_confianza_pct = _to_decimal(data.get("min_confianza_pct"), default=Decimal("0"))
        top = int(data.get("top") or 120)
        escenario = str(data.get("escenario") or "base").lower()
        qty_key = self._ESCENARIO_TO_KEY.get(escenario, "forecast_qty")
        replace_existing = bool(data.get("replace_existing", True))
        fuente = (data.get("fuente") or "API_FORECAST_STAT").strip()[:40] or "API_FORECAST_STAT"

        sucursal = None
        sucursal_id = data.get("sucursal_id")
        if sucursal_id is not None:
            sucursal = Sucursal.objects.filter(pk=sucursal_id, activa=True).first()
            if sucursal is None:
                return Response(
                    {"detail": "Sucursal no encontrada o inactiva."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        result = _build_forecast_from_history(
            alcance=alcance,
            periodo=periodo,
            fecha_base=fecha_base,
            sucursal=sucursal,
            incluir_preparaciones=incluir_preparaciones,
            safety_pct=safety_pct,
        )
        result, filtered_conf = _filter_forecast_result_by_confianza(result, min_confianza_pct)
        result["min_confianza_pct"] = min_confianza_pct
        if not result.get("rows"):
            return Response(
                {"detail": "No hay historial suficiente para generar forecast en ese alcance/filtro."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        payload = _forecast_session_payload(result, top_rows=top)
        recetas_map = {r.id: r for r in Receta.objects.filter(id__in=[int(x["receta_id"]) for x in result["rows"]]).only("id", "nombre")}

        created = 0
        updated = 0
        skipped_existing = 0
        skipped_invalid = 0
        applied_rows: list[dict[str, Any]] = []

        with transaction.atomic():
            for row in result["rows"]:
                receta_id = int(row["receta_id"])
                receta = recetas_map.get(receta_id)
                if receta is None:
                    skipped_invalid += 1
                    continue

                qty = _to_decimal(row.get(qty_key))
                if qty <= 0:
                    skipped_invalid += 1
                    continue
                qty = qty.quantize(Decimal("0.001"))

                existing = PronosticoVenta.objects.filter(receta=receta, periodo=result["periodo"]).first()
                old_qty = _to_decimal(existing.cantidad if existing else 0).quantize(Decimal("0.001"))
                if existing is None:
                    PronosticoVenta.objects.create(
                        receta=receta,
                        periodo=result["periodo"],
                        cantidad=qty,
                        fuente=fuente,
                    )
                    created += 1
                    action = "create"
                else:
                    if not replace_existing:
                        skipped_existing += 1
                        continue
                    existing.cantidad = qty
                    existing.fuente = fuente
                    existing.save(update_fields=["cantidad", "fuente", "actualizado_en"])
                    updated += 1
                    action = "update"

                if len(applied_rows) < top:
                    applied_rows.append(
                        {
                            "receta_id": receta.id,
                            "receta": receta.nombre,
                            "escenario": escenario,
                            "cantidad_anterior": _to_float(old_qty),
                            "cantidad_nueva": _to_float(qty),
                            "accion": action,
                        }
                    )

        response_payload = {
            "scope": {
                "alcance": payload["alcance"],
                "periodo": payload["periodo"],
                "target_start": payload["target_start"],
                "target_end": payload["target_end"],
                "sucursal_nombre": payload["sucursal_nombre"],
                "sucursal_id": payload.get("sucursal_id"),
                "escenario": escenario,
                "qty_key": qty_key,
                "min_confianza_pct": _to_float(min_confianza_pct),
                "filtered_conf": filtered_conf,
            },
            "totals": payload["totals"],
            "persisted": {
                "created": created,
                "updated": updated,
                "skipped_existing": skipped_existing,
                "skipped_invalid": skipped_invalid,
                "applied": created + updated,
            },
            "rows": payload["rows"],
            "applied_rows": applied_rows,
        }
        if export_format:
            return _forecast_estadistico_guardar_export_response(response_payload, export_format)
        return Response(response_payload, status=status.HTTP_200_OK)


class SolicitudVentaUpsertView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para capturar solicitudes de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = SolicitudVentaUpsertSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        receta = get_object_or_404(Receta, pk=data["receta_id"])
        sucursal = None
        if data.get("sucursal_id") is not None:
            sucursal = Sucursal.objects.filter(pk=data["sucursal_id"], activa=True).first()
            if sucursal is None:
                return Response(
                    {"detail": "Sucursal no encontrada o inactiva."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        alcance = _ui_to_model_alcance(data.get("alcance"))
        periodo_default = _normalize_periodo_mes(data.get("periodo"))
        fecha_base_default = data.get("fecha_base") or timezone.localdate()
        periodo, fecha_inicio, fecha_fin = _resolve_solicitud_window(
            alcance=alcance,
            periodo_raw=data.get("periodo"),
            fecha_base_raw=data.get("fecha_base"),
            fecha_inicio_raw=data.get("fecha_inicio"),
            fecha_fin_raw=data.get("fecha_fin"),
            periodo_default=periodo_default,
            fecha_base_default=fecha_base_default,
        )
        cantidad = _to_decimal(data.get("cantidad"))
        fuente = (data.get("fuente") or "API_SOL_VENTAS").strip()[:40] or "API_SOL_VENTAS"

        with transaction.atomic():
            record, created = SolicitudVenta.objects.get_or_create(
                receta=receta,
                sucursal=sucursal,
                alcance=alcance,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                defaults={
                    "periodo": periodo,
                    "cantidad": cantidad,
                    "fuente": fuente,
                },
            )
            if not created:
                record.periodo = periodo
                record.cantidad = cantidad
                record.fuente = fuente
                record.save(update_fields=["periodo", "cantidad", "fuente", "actualizado_en"])

        forecast_ref = None
        validate_forecast = _parse_bool(request.GET.get("validate_forecast"), default=True)
        if validate_forecast:
            forecast_alcance = {
                SolicitudVenta.ALCANCE_MES: "mes",
                SolicitudVenta.ALCANCE_SEMANA: "semana",
                SolicitudVenta.ALCANCE_FIN_SEMANA: "fin_semana",
            }.get(alcance, "mes")
            forecast_incluir_preparaciones = _parse_bool(request.GET.get("incluir_preparaciones"), default=True)
            forecast_safety_pct = _to_decimal(request.GET.get("safety_pct"), default=Decimal("0"))
            forecast_min_confianza_pct = _to_decimal(request.GET.get("min_confianza_pct"), default=Decimal("0"))
            threshold_pct = _to_decimal(request.GET.get("umbral_desviacion_pct"), default=Decimal("20"))
            if threshold_pct < 0:
                threshold_pct = Decimal("20")

            try:
                forecast_result = _build_forecast_from_history(
                    alcance=forecast_alcance,
                    periodo=periodo,
                    fecha_base=fecha_inicio,
                    sucursal=sucursal,
                    incluir_preparaciones=forecast_incluir_preparaciones,
                    safety_pct=forecast_safety_pct,
                )
                forecast_result, _ = _filter_forecast_result_by_confianza(forecast_result, forecast_min_confianza_pct)
                row = next(
                    (r for r in (forecast_result.get("rows") or []) if int(r.get("receta_id") or 0) == receta.id),
                    None,
                )
                if row:
                    forecast_qty = _to_decimal(row.get("forecast_qty"))
                    delta_qty = (cantidad - forecast_qty).quantize(Decimal("0.001"))
                    variacion_pct = None
                    if forecast_qty > 0:
                        variacion_pct = ((delta_qty / forecast_qty) * Decimal("100")).quantize(Decimal("0.1"))

                    if forecast_qty <= 0 and cantidad > 0:
                        status_forecast = "SIN_BASE"
                    elif variacion_pct is None:
                        status_forecast = "OK"
                    elif variacion_pct > threshold_pct:
                        status_forecast = "SOBRE"
                    elif variacion_pct < (-threshold_pct):
                        status_forecast = "BAJO"
                    else:
                        status_forecast = "OK"

                    forecast_ref = {
                        "status": status_forecast,
                        "threshold_pct": _to_float(threshold_pct),
                        "forecast_qty": _to_float(forecast_qty),
                        "forecast_low": _to_float(row.get("forecast_low")),
                        "forecast_high": _to_float(row.get("forecast_high")),
                        "solicitud_qty": _to_float(cantidad),
                        "delta_qty": _to_float(delta_qty),
                        "variacion_pct": _to_float(variacion_pct) if variacion_pct is not None else None,
                        "confianza_pct": _to_float(row.get("confianza")),
                        "recomendacion": row.get("recomendacion") or "",
                    }
                else:
                    forecast_ref = {"status": "SIN_FORECAST"}
            except Exception:
                forecast_ref = {"status": "FORECAST_ERROR"}

        return Response(
            {
                "created": created,
                "id": record.id,
                "receta_id": record.receta_id,
                "receta": record.receta.nombre,
                "sucursal_id": record.sucursal_id,
                "sucursal": record.sucursal.nombre if record.sucursal_id else "",
                "alcance": record.alcance,
                "periodo": record.periodo,
                "fecha_inicio": str(record.fecha_inicio),
                "fecha_fin": str(record.fecha_fin),
                "cantidad": str(record.cantidad),
                "fuente": record.fuente,
                "forecast_ref": forecast_ref,
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )


class SolicitudVentaAplicarForecastView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para ajustar solicitudes de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )
        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ser = SolicitudVentaAplicarForecastSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        sucursal = Sucursal.objects.filter(pk=data["sucursal_id"], activa=True).first()
        if sucursal is None:
            return Response(
                {"detail": "Sucursal no encontrada o inactiva."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        alcance = data.get("alcance") or "mes"
        periodo = _normalize_periodo_mes(data.get("periodo"))
        fecha_base = data.get("fecha_base") or timezone.localdate()
        incluir_preparaciones = bool(data.get("incluir_preparaciones"))
        safety_pct = _to_decimal(data.get("safety_pct"), default=Decimal("0"))
        min_confianza_pct = _to_decimal(data.get("min_confianza_pct"), default=Decimal("0"))
        escenario = str(data.get("escenario") or "base").lower()
        dry_run = bool(data.get("dry_run", False))
        max_variacion_pct = _to_decimal(data.get("max_variacion_pct"), default=Decimal("-1"))
        if max_variacion_pct < 0:
            max_variacion_pct = None
        top = int(data.get("top") or 120)

        result = _build_forecast_from_history(
            alcance=alcance,
            periodo=periodo,
            fecha_base=fecha_base,
            sucursal=sucursal,
            incluir_preparaciones=incluir_preparaciones,
            safety_pct=safety_pct,
        )
        result, _ = _filter_forecast_result_by_confianza(result, min_confianza_pct)
        if not result.get("rows"):
            return Response(
                {"detail": "No hay forecast tras aplicar el filtro de confianza mínima."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        full_payload = _forecast_session_payload(result, top_rows=max(len(result.get("rows") or []), 1))
        compare_raw = _forecast_vs_solicitud_preview(full_payload, escenario=escenario)
        if not compare_raw or not compare_raw.get("rows"):
            return Response(
                {"detail": "No hay filas de comparación forecast vs solicitud para aplicar ajuste."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        modo = data.get("modo") or "desviadas"
        receta_id = data.get("receta_id")
        rows = list(compare_raw["rows"])
        if modo == "sobre":
            rows = [r for r in rows if r.get("status") == "SOBRE"]
        elif modo == "bajo":
            rows = [r for r in rows if r.get("status") == "BAJO"]
        elif modo == "todas":
            rows = [r for r in rows if r.get("status") in {"SOBRE", "BAJO", "SIN_BASE", "OK"}]
        elif modo == "receta":
            rows = [r for r in rows if int(r.get("receta_id") or 0) == int(receta_id)]
        else:
            rows = [r for r in rows if r.get("status") in {"SOBRE", "BAJO"}]

        if not rows:
            return Response(
                {"detail": "No hay filas objetivo para el modo seleccionado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        model_alcance = _ui_to_model_alcance(alcance)
        target_start = result["target_start"]
        target_end = result["target_end"]
        result_periodo = result["periodo"]
        fuente = (data.get("fuente") or "API_FORECAST_ADJUST").strip()[:40] or "API_FORECAST_ADJUST"

        created = 0
        updated = 0
        skipped = 0
        skipped_cap = 0
        applied = 0
        adjusted_rows = []
        tx_cm = nullcontext() if dry_run else transaction.atomic()
        with tx_cm:
            for row in rows:
                receta = Receta.objects.filter(pk=row["receta_id"]).first()
                if receta is None:
                    skipped += 1
                    continue
                nueva_cantidad = _to_decimal(row.get("forecast_qty"))
                if nueva_cantidad < 0:
                    skipped += 1
                    continue
                record = SolicitudVenta.objects.filter(
                    receta=receta,
                    sucursal=sucursal,
                    alcance=model_alcance,
                    fecha_inicio=target_start,
                    fecha_fin=target_end,
                ).first()
                was_created = record is None
                old_qty = _to_decimal(record.cantidad if record is not None else 0)
                variacion_pct = None
                if old_qty > 0:
                    variacion_pct = ((nueva_cantidad - old_qty) / old_qty * Decimal("100")).quantize(Decimal("0.1"))
                if (max_variacion_pct is not None) and (variacion_pct is not None):
                    if abs(variacion_pct) > max_variacion_pct:
                        skipped += 1
                        skipped_cap += 1
                        continue
                if was_created:
                    if dry_run:
                        created += 1
                    else:
                        SolicitudVenta.objects.create(
                            receta=receta,
                            sucursal=sucursal,
                            alcance=model_alcance,
                            periodo=result_periodo,
                            fecha_inicio=target_start,
                            fecha_fin=target_end,
                            cantidad=nueva_cantidad,
                            fuente=fuente,
                        )
                        created += 1
                        applied += 1
                else:
                    if dry_run:
                        updated += 1
                    else:
                        record.periodo = result_periodo
                        record.cantidad = nueva_cantidad
                        record.fuente = fuente
                        record.save(update_fields=["periodo", "cantidad", "fuente", "actualizado_en"])
                        updated += 1
                        applied += 1

                adjusted_rows.append(
                    {
                        "receta_id": receta.id,
                        "receta": receta.nombre,
                        "anterior": _to_float(old_qty),
                        "nueva": _to_float(nueva_cantidad),
                        "variacion_pct": _to_float(variacion_pct) if variacion_pct is not None else None,
                        "accion": "create" if was_created else "update",
                        "status_before": row.get("status") or "",
                    }
                )

        compare_payload = _serialize_forecast_compare(compare_raw, top=top)
        response_payload = {
            "scope": {
                "alcance": alcance,
                "periodo": result_periodo,
                "target_start": str(target_start),
                "target_end": str(target_end),
                "sucursal_id": sucursal.id,
                "sucursal_nombre": f"{sucursal.codigo} - {sucursal.nombre}",
                "modo": modo,
                "escenario": escenario,
                "min_confianza_pct": _to_float(min_confianza_pct),
            },
            "updated": {
                "dry_run": dry_run,
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "skipped_cap": skipped_cap,
                "applied": applied,
            },
            "adjusted_rows": adjusted_rows[:top],
            "compare_solicitud": compare_payload,
        }
        if export_format:
            return _ventas_solicitud_aplicar_forecast_export_response(response_payload, export_format)
        return Response(response_payload, status=status.HTTP_200_OK)


class ActivosCalendarioMantenimientoView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not can_view_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar calendario de mantenimiento."},
                status=status.HTTP_403_FORBIDDEN,
            )

        today = timezone.localdate()
        date_from = _parse_iso_date(request.GET.get("from")) or today
        date_to = _parse_iso_date(request.GET.get("to")) or (date_from + timedelta(days=45))
        if date_to < date_from:
            date_to = date_from + timedelta(days=45)

        planes = list(
            PlanMantenimiento.objects.select_related("activo_ref")
            .filter(
                activo=True,
                estatus=PlanMantenimiento.ESTATUS_ACTIVO,
                proxima_ejecucion__isnull=False,
                proxima_ejecucion__gte=date_from,
                proxima_ejecucion__lte=date_to,
            )
            .order_by("proxima_ejecucion", "id")
        )
        ordenes = list(
            OrdenMantenimiento.objects.select_related("activo_ref", "plan_ref")
            .filter(
                fecha_programada__gte=date_from,
                fecha_programada__lte=date_to,
            )
            .order_by("fecha_programada", "id")
        )

        events = []
        for plan in planes:
            events.append(
                {
                    "fecha": str(plan.proxima_ejecucion),
                    "tipo": "PLAN",
                    "referencia": f"PLAN-{plan.id}",
                    "activo_id": plan.activo_ref_id,
                    "activo": plan.activo_ref.nombre,
                    "detalle": plan.nombre,
                    "estado": plan.estatus,
                    "responsable": plan.responsable or "",
                }
            )
        for orden in ordenes:
            events.append(
                {
                    "fecha": str(orden.fecha_programada),
                    "tipo": "ORDEN",
                    "referencia": orden.folio,
                    "activo_id": orden.activo_ref_id,
                    "activo": orden.activo_ref.nombre,
                    "detalle": orden.descripcion or orden.get_tipo_display(),
                    "estado": orden.estatus,
                    "responsable": orden.responsable or "",
                }
            )
        events.sort(key=lambda row: (row["fecha"], row["tipo"], row["referencia"]))

        return Response(
            {
                "range": {
                    "from": str(date_from),
                    "to": str(date_to),
                    "days": (date_to - date_from).days + 1,
                },
                "totales": {
                    "planes": len(planes),
                    "ordenes": len(ordenes),
                    "eventos": len(events),
                },
                "events": events,
            },
            status=status.HTTP_200_OK,
        )


class ActivosDisponibilidadView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not can_view_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar disponibilidad de activos."},
                status=status.HTTP_403_FORBIDDEN,
            )

        activos_qs = Activo.objects.filter(activo=True)
        total = activos_qs.count()
        operativos = activos_qs.filter(estado=Activo.ESTADO_OPERATIVO).count()
        en_mantenimiento = activos_qs.filter(estado=Activo.ESTADO_MANTENIMIENTO).count()
        fuera_servicio = activos_qs.filter(estado=Activo.ESTADO_FUERA_SERVICIO).count()
        disponibilidad_pct = round((operativos * 100.0 / total), 2) if total else 100.0

        criticidad_rows = (
            activos_qs.values("criticidad")
            .annotate(total=Count("id"))
            .order_by("criticidad")
        )
        criticidad = {row["criticidad"]: int(row["total"] or 0) for row in criticidad_rows}

        hoy = timezone.localdate()
        ordenes_abiertas = OrdenMantenimiento.objects.filter(
            estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO]
        )
        planes_vencidos = PlanMantenimiento.objects.filter(
            activo=True,
            estatus=PlanMantenimiento.ESTATUS_ACTIVO,
            proxima_ejecucion__isnull=False,
            proxima_ejecucion__lt=hoy,
        ).count()

        return Response(
            {
                "totales": {
                    "activos": total,
                    "operativos": operativos,
                    "en_mantenimiento": en_mantenimiento,
                    "fuera_servicio": fuera_servicio,
                    "disponibilidad_pct": disponibilidad_pct,
                    "ordenes_abiertas": ordenes_abiertas.count(),
                    "ordenes_en_proceso": ordenes_abiertas.filter(estatus=OrdenMantenimiento.ESTATUS_EN_PROCESO).count(),
                    "planes_vencidos": planes_vencidos,
                },
                "criticidad": {
                    "ALTA": criticidad.get(Activo.CRITICIDAD_ALTA, 0),
                    "MEDIA": criticidad.get(Activo.CRITICIDAD_MEDIA, 0),
                    "BAJA": criticidad.get(Activo.CRITICIDAD_BAJA, 0),
                },
            },
            status=status.HTTP_200_OK,
        )


class ActivosOrdenesView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not can_manage_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para crear órdenes de mantenimiento."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = ActivosOrdenCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        activo = get_object_or_404(Activo, pk=data["activo_id"], activo=True)
        plan = None
        plan_id = data.get("plan_id")
        if plan_id:
            plan = get_object_or_404(PlanMantenimiento, pk=plan_id, activo_ref=activo)

        fecha_programada = data.get("fecha_programada") or timezone.localdate()
        orden = OrdenMantenimiento.objects.create(
            activo_ref=activo,
            plan_ref=plan,
            tipo=data.get("tipo") or OrdenMantenimiento.TIPO_PREVENTIVO,
            prioridad=data.get("prioridad") or OrdenMantenimiento.PRIORIDAD_MEDIA,
            fecha_programada=fecha_programada,
            responsable=(data.get("responsable") or "").strip(),
            descripcion=(data.get("descripcion") or "").strip(),
            creado_por=request.user,
        )
        BitacoraMantenimiento.objects.create(
            orden=orden,
            accion="CREADA",
            comentario="Orden creada desde API",
            usuario=request.user,
        )
        log_event(
            request.user,
            "CREATE",
            "activos.OrdenMantenimiento",
            orden.id,
            {
                "folio": orden.folio,
                "activo_id": orden.activo_ref_id,
                "plan_id": orden.plan_ref_id,
                "tipo": orden.tipo,
                "prioridad": orden.prioridad,
                "estatus": orden.estatus,
                "source": "api",
            },
        )

        return Response(
            {
                "id": orden.id,
                "folio": orden.folio,
                "activo_id": orden.activo_ref_id,
                "activo": orden.activo_ref.nombre,
                "plan_id": orden.plan_ref_id,
                "tipo": orden.tipo,
                "prioridad": orden.prioridad,
                "estatus": orden.estatus,
                "fecha_programada": str(orden.fecha_programada),
                "responsable": orden.responsable,
                "descripcion": orden.descripcion,
            },
            status=status.HTTP_201_CREATED,
        )


class ActivosOrdenStatusUpdateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, orden_id: int):
        if not can_manage_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para actualizar estatus de órdenes de mantenimiento."},
                status=status.HTTP_403_FORBIDDEN,
            )

        orden = get_object_or_404(OrdenMantenimiento.objects.select_related("plan_ref"), pk=orden_id)
        ser = ActivosOrdenStatusSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        estatus_new = ser.validated_data["estatus"]
        estatus_prev = orden.estatus
        if estatus_prev == estatus_new:
            return Response(
                {
                    "id": orden.id,
                    "folio": orden.folio,
                    "from": estatus_prev,
                    "to": estatus_new,
                    "updated": False,
                },
                status=status.HTTP_200_OK,
            )

        allowed_transitions = {
            OrdenMantenimiento.ESTATUS_PENDIENTE: {
                OrdenMantenimiento.ESTATUS_EN_PROCESO,
                OrdenMantenimiento.ESTATUS_CERRADA,
                OrdenMantenimiento.ESTATUS_CANCELADA,
            },
            OrdenMantenimiento.ESTATUS_EN_PROCESO: {
                OrdenMantenimiento.ESTATUS_CERRADA,
                OrdenMantenimiento.ESTATUS_CANCELADA,
            },
            OrdenMantenimiento.ESTATUS_CERRADA: set(),
            OrdenMantenimiento.ESTATUS_CANCELADA: set(),
        }
        if estatus_new not in allowed_transitions.get(estatus_prev, set()):
            return Response(
                {"detail": f"Transición inválida: {estatus_prev} -> {estatus_new}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        today = timezone.localdate()
        orden.estatus = estatus_new
        update_fields = ["estatus", "actualizado_en"]
        if estatus_new == OrdenMantenimiento.ESTATUS_EN_PROCESO and not orden.fecha_inicio:
            orden.fecha_inicio = today
            update_fields.append("fecha_inicio")
        if estatus_new == OrdenMantenimiento.ESTATUS_CERRADA:
            orden.fecha_cierre = today
            update_fields.append("fecha_cierre")
            if orden.plan_ref_id:
                plan = orden.plan_ref
                plan.ultima_ejecucion = today
                plan.recompute_next_date()
                plan.save(update_fields=["ultima_ejecucion", "proxima_ejecucion", "actualizado_en"])
        orden.save(update_fields=update_fields)

        BitacoraMantenimiento.objects.create(
            orden=orden,
            accion="ESTATUS",
            comentario=f"{estatus_prev} -> {estatus_new}",
            usuario=request.user,
        )
        log_event(
            request.user,
            "UPDATE",
            "activos.OrdenMantenimiento",
            orden.id,
            {"from": estatus_prev, "to": estatus_new, "folio": orden.folio, "source": "api"},
        )

        return Response(
            {
                "id": orden.id,
                "folio": orden.folio,
                "from": estatus_prev,
                "to": estatus_new,
                "updated": True,
                "fecha_inicio": str(orden.fecha_inicio) if orden.fecha_inicio else None,
                "fecha_cierre": str(orden.fecha_cierre) if orden.fecha_cierre else None,
            },
            status=status.HTTP_200_OK,
        )


def _resolve_recipe_for_pos_row(*, row: dict[str, Any], receta_cache: dict[tuple[int, str, str], Receta | None]) -> Receta | None:
    receta_id = row.get("receta_id")
    receta_name = str(row.get("receta") or row.get("producto") or "").strip()
    codigo_point = str(row.get("codigo_point") or "").strip()
    return _resolve_receta_bulk_ref(
        receta_id=receta_id,
        receta_name=receta_name,
        codigo_point=codigo_point,
        cache=receta_cache,
    )


def _resolve_sucursal_for_pos_row(
    *,
    row: dict[str, Any],
    default_sucursal: Sucursal | None,
    sucursal_cache: dict[tuple[int, str, str, int], Sucursal | None],
) -> Sucursal | None:
    return _resolve_sucursal_bulk_ref(
        sucursal_id=row.get("sucursal_id"),
        sucursal_name=str(row.get("sucursal") or "").strip(),
        sucursal_codigo=str(row.get("sucursal_codigo") or "").strip(),
        default_sucursal=default_sucursal,
        cache=sucursal_cache,
    )


def _process_control_venta_pos_bulk(
    data: dict[str, Any],
    *,
    dry_run_override: bool | None = None,
) -> dict[str, Any]:
    rows = data["rows"]
    modo = data.get("modo") or "replace"
    fuente = (data.get("fuente") or "API_POS_VENTAS").strip()[:40] or "API_POS_VENTAS"
    dry_run = bool(data.get("dry_run", True) if dry_run_override is None else dry_run_override)
    stop_on_error = bool(data.get("stop_on_error", False))
    top = int(data.get("top") or 120)

    default_sucursal = None
    default_sucursal_id = data.get("sucursal_default_id")
    if default_sucursal_id is not None:
        default_sucursal = Sucursal.objects.filter(pk=default_sucursal_id, activa=True).first()
        if default_sucursal is None:
            raise ValueError("Sucursal default no encontrada o inactiva.")

    receta_cache: dict[tuple[int, str, str], Receta | None] = {}
    sucursal_cache: dict[tuple[int, str, str, int], Sucursal | None] = {}
    created = 0
    updated = 0
    skipped = 0
    terminated_early = False
    result_rows: list[dict[str, Any]] = []

    tx_cm = nullcontext() if dry_run else transaction.atomic()
    with tx_cm:
        for index, row in enumerate(rows, start=1):
            receta = _resolve_recipe_for_pos_row(row=row, receta_cache=receta_cache)
            producto_texto = str(row.get("producto") or row.get("receta") or "").strip()
            codigo_point = str(row.get("codigo_point") or "").strip()
            if receta is None and not (producto_texto or codigo_point):
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "producto_not_found",
                        "producto_input": producto_texto or codigo_point,
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            sucursal = _resolve_sucursal_for_pos_row(
                row=row,
                default_sucursal=default_sucursal,
                sucursal_cache=sucursal_cache,
            )
            has_sucursal_ref = bool(row.get("sucursal_id")) or bool(row.get("sucursal")) or bool(row.get("sucursal_codigo"))
            if has_sucursal_ref and sucursal is None:
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "sucursal_not_found",
                        "sucursal_input": row.get("sucursal_codigo") or row.get("sucursal") or row.get("sucursal_id"),
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            fecha = row["fecha"]
            cantidad = _to_decimal(row.get("cantidad"), default=Decimal("0"))
            tickets = int(row.get("tickets") or 0)
            monto_total_raw = row.get("monto_total")
            monto_total = _to_decimal(monto_total_raw) if monto_total_raw is not None else None

            existing_qs = VentaPOS.objects.filter(
                receta=receta,
                sucursal=sucursal,
                fecha=fecha,
                codigo_point=codigo_point,
                producto_texto=producto_texto,
            )
            existing = existing_qs.order_by("id").first()

            previous_qty = _to_decimal(existing.cantidad, default=Decimal("0")) if existing else Decimal("0")
            previous_tickets = int(existing.tickets or 0) if existing else 0
            if existing:
                if modo == "accumulate":
                    new_qty = previous_qty + cantidad
                    new_tickets = previous_tickets + tickets
                    new_monto = (_to_decimal(existing.monto_total, default=Decimal("0")) + _to_decimal(monto_total, default=Decimal("0")))
                else:
                    new_qty = cantidad
                    new_tickets = tickets
                    new_monto = monto_total
                action = "UPDATED"
                updated += 1
            else:
                new_qty = cantidad
                new_tickets = tickets
                new_monto = monto_total
                action = "CREATED"
                created += 1

            if not dry_run:
                if existing:
                    existing.cantidad = new_qty
                    existing.tickets = new_tickets
                    existing.monto_total = new_monto
                    existing.fuente = fuente
                    existing.save(update_fields=["cantidad", "tickets", "monto_total", "fuente", "actualizado_en"])
                else:
                    VentaPOS.objects.create(
                        receta=receta,
                        sucursal=sucursal,
                        fecha=fecha,
                        codigo_point=codigo_point,
                        producto_texto=producto_texto,
                        cantidad=new_qty,
                        tickets=new_tickets,
                        monto_total=new_monto,
                        fuente=fuente,
                    )

            result_rows.append(
                {
                    "row": index,
                    "status": action,
                    "receta_id": receta.id if receta else None,
                    "receta": receta.nombre if receta else "",
                    "producto": producto_texto,
                    "codigo_point": codigo_point,
                    "sucursal_id": sucursal.id if sucursal else None,
                    "sucursal": sucursal.nombre if sucursal else "",
                    "fecha": str(fecha),
                    "cantidad_prev": float(previous_qty),
                    "cantidad_nueva": float(new_qty),
                    "tickets_prev": previous_tickets,
                    "tickets_nuevos": new_tickets,
                }
            )

    error_count = sum(1 for row in result_rows if row.get("status") == "ERROR")
    return {
        "dry_run": dry_run,
        "mode": modo,
        "fuente": fuente,
        "terminated_early": terminated_early,
        "summary": {
            "total_rows": len(rows),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": error_count,
            "applied": 0 if dry_run else (created + updated),
        },
        "rows": result_rows[:top],
    }


def _process_control_merma_pos_bulk(
    data: dict[str, Any],
    *,
    dry_run_override: bool | None = None,
) -> dict[str, Any]:
    rows = data["rows"]
    modo = data.get("modo") or "replace"
    fuente = (data.get("fuente") or "API_POS_MERMAS").strip()[:40] or "API_POS_MERMAS"
    dry_run = bool(data.get("dry_run", True) if dry_run_override is None else dry_run_override)
    stop_on_error = bool(data.get("stop_on_error", False))
    top = int(data.get("top") or 120)

    default_sucursal = None
    default_sucursal_id = data.get("sucursal_default_id")
    if default_sucursal_id is not None:
        default_sucursal = Sucursal.objects.filter(pk=default_sucursal_id, activa=True).first()
        if default_sucursal is None:
            raise ValueError("Sucursal default no encontrada o inactiva.")

    receta_cache: dict[tuple[int, str, str], Receta | None] = {}
    sucursal_cache: dict[tuple[int, str, str, int], Sucursal | None] = {}
    created = 0
    updated = 0
    skipped = 0
    terminated_early = False
    result_rows: list[dict[str, Any]] = []

    tx_cm = nullcontext() if dry_run else transaction.atomic()
    with tx_cm:
        for index, row in enumerate(rows, start=1):
            receta = _resolve_recipe_for_pos_row(row=row, receta_cache=receta_cache)
            producto_texto = str(row.get("producto") or row.get("receta") or "").strip()
            codigo_point = str(row.get("codigo_point") or "").strip()
            if receta is None and not (producto_texto or codigo_point):
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "producto_not_found",
                        "producto_input": producto_texto or codigo_point,
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            sucursal = _resolve_sucursal_for_pos_row(
                row=row,
                default_sucursal=default_sucursal,
                sucursal_cache=sucursal_cache,
            )
            has_sucursal_ref = bool(row.get("sucursal_id")) or bool(row.get("sucursal")) or bool(row.get("sucursal_codigo"))
            if has_sucursal_ref and sucursal is None:
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "sucursal_not_found",
                        "sucursal_input": row.get("sucursal_codigo") or row.get("sucursal") or row.get("sucursal_id"),
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            fecha = row["fecha"]
            cantidad = _to_decimal(row.get("cantidad"), default=Decimal("0"))
            motivo = str(row.get("motivo") or "").strip()[:160]

            existing_qs = MermaPOS.objects.filter(
                receta=receta,
                sucursal=sucursal,
                fecha=fecha,
                codigo_point=codigo_point,
                producto_texto=producto_texto,
                motivo=motivo,
            )
            existing = existing_qs.order_by("id").first()

            previous_qty = _to_decimal(existing.cantidad, default=Decimal("0")) if existing else Decimal("0")
            if existing:
                new_qty = previous_qty + cantidad if modo == "accumulate" else cantidad
                action = "UPDATED"
                updated += 1
            else:
                new_qty = cantidad
                action = "CREATED"
                created += 1

            if not dry_run:
                if existing:
                    existing.cantidad = new_qty
                    existing.fuente = fuente
                    existing.save(update_fields=["cantidad", "fuente", "actualizado_en"])
                else:
                    MermaPOS.objects.create(
                        receta=receta,
                        sucursal=sucursal,
                        fecha=fecha,
                        codigo_point=codigo_point,
                        producto_texto=producto_texto,
                        cantidad=new_qty,
                        motivo=motivo,
                        fuente=fuente,
                    )

            result_rows.append(
                {
                    "row": index,
                    "status": action,
                    "receta_id": receta.id if receta else None,
                    "receta": receta.nombre if receta else "",
                    "producto": producto_texto,
                    "codigo_point": codigo_point,
                    "sucursal_id": sucursal.id if sucursal else None,
                    "sucursal": sucursal.nombre if sucursal else "",
                    "fecha": str(fecha),
                    "cantidad_prev": float(previous_qty),
                    "cantidad_nueva": float(new_qty),
                }
            )

    error_count = sum(1 for row in result_rows if row.get("status") == "ERROR")
    return {
        "dry_run": dry_run,
        "mode": modo,
        "fuente": fuente,
        "terminated_early": terminated_early,
        "summary": {
            "total_rows": len(rows),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": error_count,
            "applied": 0 if dry_run else (created + updated),
        },
        "rows": result_rows[:top],
    }


class ControlVentasPosImportPreviewView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not (can_manage_inventario(request.user) or can_manage_compras(request.user)):
            return Response(
                {"detail": "No tienes permisos para previsualizar importación POS."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = ControlVentaPosBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_control_venta_pos_bulk(ser.validated_data, dry_run_override=True)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        payload["preview"] = True
        return Response(payload, status=status.HTTP_200_OK)


class ControlVentasPosImportConfirmView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not (can_manage_inventario(request.user) or can_manage_compras(request.user)):
            return Response(
                {"detail": "No tienes permisos para confirmar importación POS."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = ControlVentaPosBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_control_venta_pos_bulk(ser.validated_data, dry_run_override=False)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        payload["preview"] = False
        return Response(payload, status=status.HTTP_200_OK)


class ControlVentasPosBulkUpsertView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not (can_manage_inventario(request.user) or can_manage_compras(request.user)):
            return Response(
                {"detail": "No tienes permisos para importar ventas POS."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = ControlVentaPosBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_control_venta_pos_bulk(ser.validated_data)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(payload, status=status.HTTP_200_OK)


class ControlMermasPosImportPreviewView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not (can_manage_inventario(request.user) or can_manage_compras(request.user)):
            return Response(
                {"detail": "No tienes permisos para previsualizar importación de mermas POS."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = ControlMermaPosBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_control_merma_pos_bulk(ser.validated_data, dry_run_override=True)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        payload["preview"] = True
        return Response(payload, status=status.HTTP_200_OK)


class ControlMermasPosImportConfirmView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not (can_manage_inventario(request.user) or can_manage_compras(request.user)):
            return Response(
                {"detail": "No tienes permisos para confirmar importación de mermas POS."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = ControlMermaPosBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_control_merma_pos_bulk(ser.validated_data, dry_run_override=False)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        payload["preview"] = False
        return Response(payload, status=status.HTTP_200_OK)


class ControlMermasPosBulkUpsertView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not (can_manage_inventario(request.user) or can_manage_compras(request.user)):
            return Response(
                {"detail": "No tienes permisos para importar mermas POS."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = ControlMermaPosBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_control_merma_pos_bulk(ser.validated_data)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(payload, status=status.HTTP_200_OK)


class ControlDiscrepanciasView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not can_view_reportes(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar discrepancias."},
                status=status.HTTP_403_FORBIDDEN,
            )

        date_from, date_to, period_resolved = resolve_period_range(
            period_raw=request.GET.get("periodo"),
            date_from_raw=request.GET.get("from"),
            date_to_raw=request.GET.get("to"),
        )
        sucursal_id = None
        sucursal_id_raw = (request.GET.get("sucursal_id") or "").strip()
        if sucursal_id_raw.isdigit():
            sucursal_id = int(sucursal_id_raw)
        threshold = _to_decimal(request.GET.get("threshold_pct"), default=Decimal("10"))
        top = _parse_bounded_int(request.GET.get("top"), default=300, min_value=1, max_value=1000)

        report = build_discrepancias_report(
            date_from=date_from,
            date_to=date_to,
            sucursal_id=sucursal_id,
            threshold_pct=threshold,
            top=top,
        )
        report["scope"] = {
            "periodo": period_resolved,
            "sucursal_id": sucursal_id,
            "top": top,
        }
        return Response(report, status=status.HTTP_200_OK)
