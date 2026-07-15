import csv
import json
from collections import defaultdict
from contextlib import nullcontext
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from django.db import transaction, OperationalError, ProgrammingError
from django.db.models import Count, Max, Q, Sum
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from openpyxl import Workbook
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.authtoken.models import Token
from rest_framework.authtoken.views import ObtainAuthToken

from activos.models import Activo, BitacoraMantenimiento, OrdenMantenimiento, PlanMantenimiento
from control.models import MermaPOS, VentaPOS
from control.services import build_discrepancias_report, resolve_period_range
from compras.models import OrdenCompra, RecepcionCompra, SolicitudCompra
from core.access import (
    ROLE_ADMIN,
    ROLE_COMPRAS,
    ROLE_DG,
    can_manage_compras,
    can_manage_inventario,
    can_view_audit,
    can_view_compras,
    can_view_inventario,
    can_view_maestros,
    can_view_reportes,
    has_any_role,
)
from core.audit import log_event
from core.models import AuditLog, Sucursal
from compras.views import (
    _apply_recepcion_to_inventario,
    _active_solicitud_statuses,
    _build_budget_context,
    _build_budget_history,
    _build_category_dashboard,
    _can_transition_orden,
    _can_transition_recepcion,
    _can_transition_solicitud,
    _build_consumo_vs_plan_dashboard,
    _default_fecha_requerida,
    _parse_date_value,
    _build_provider_dashboard,
    _filtered_solicitudes,
    _resolve_proveedor_name,
    _sanitize_consumo_ref_filter,
)
from integraciones.models import PublicApiAccessLog, PublicApiClient
from integraciones.views import _deactivate_idle_api_clients, _purge_api_logs
from inventario.models import AjusteInventario, AlmacenSyncRun, ExistenciaInsumo
from inventario.views import (
    _apply_ajuste,
    _apply_cross_filters,
    _build_cross_unified_rows,
    _export_cross_pending_csv,
    _export_cross_pending_xlsx,
    _build_pending_grouped,
    _resolve_cross_source_with_alias,
)
from maestros.models import CostoInsumo, Insumo, InsumoAlias, PointPendingMatch, Proveedor
from maestros.utils.canonical_catalog import canonical_insumo, canonical_insumo_by_id, canonicalized_active_insumos, latest_costo_canonico
from recetas.models import (
    LineaReceta,
    PlanProduccion,
    PlanProduccionItem,
    PronosticoVenta,
    Receta,
    RecetaCodigoPointAlias,
    SolicitudVenta,
    VentaHistorica,
    normalizar_codigo_point,
)
from recetas.views import (
    _build_forecast_backtest_preview,
    _build_forecast_from_history,
    _filter_forecast_result_by_confianza,
    _forecast_session_payload,
    _forecast_vs_solicitud_preview,
    _normalize_periodo_mes,
    _resolve_receta_for_sales,
    _resolve_solicitud_window,
    _resolve_sucursal_for_sales,
    _ui_to_model_alcance,
)
from recetas.utils.normalizacion import normalizar_nombre
from recetas.utils.matching import match_insumo
from recetas.utils.costeo_versionado import asegurar_version_costeo, comparativo_versiones
from ..serializers import (
    ComprasSolicitudImportConfirmSerializer,
    ComprasSolicitudImportPreviewSerializer,
    ComprasCrearOrdenSerializer,
    ComprasOrdenStatusSerializer,
    ComprasRecepcionCreateSerializer,
    ComprasRecepcionStatusSerializer,
    ComprasSolicitudCreateSerializer,
    ComprasSolicitudStatusSerializer,
    ControlMermaPosBulkSerializer,
    ControlVentaPosBulkSerializer,
    ActivosOrdenCreateSerializer,
    ActivosOrdenStatusSerializer,
    ForecastBacktestRequestSerializer,
    ForecastEstadisticoGuardarSerializer,
    ForecastEstadisticoRequestSerializer,
    IntegracionesDeactivateIdleClientsSerializer,
    IntegracionesMaintenanceRunSerializer,
    IntegracionesOperationHistoryQuerySerializer,
    IntegracionesPurgeApiLogsSerializer,
    InventarioAjusteCreateSerializer,
    InventarioAjusteDecisionSerializer,
    InventarioAliasCreateSerializer,
    InventarioCrossPendientesResolveSerializer,
    InventarioAliasMassReassignSerializer,
    MasterDuplicatesSerializer,
    MasterNormalizeSerializer,
    InventarioPointPendingResolveSerializer,
    MRPRequestSerializer,
    MRPRequerimientosRequestSerializer,
    PlanProduccionCreateSerializer,
    PlanProduccionItemCreateSerializer,
    PlanProduccionItemUpdateSerializer,
    PlanProduccionUpdateSerializer,
    PlanDesdePronosticoRequestSerializer,
    PronosticoVentaBulkSerializer,
    RecetaCostoVersionSerializer,
    SolicitudVentaAplicarForecastSerializer,
    SolicitudVentaBulkSerializer,
    SolicitudVentaUpsertSerializer,
    VentaHistoricaBulkSerializer,
)


def _load_versiones_costeo(receta: Receta, limit: int):
    return list(receta.versiones_costo.order_by("-version_num")[:limit])


def _to_decimal(value, default: Decimal = Decimal("0")) -> Decimal:
    if value is None:
        return default
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return default


def _canonical_member_ids(insumo_id: int | str | None) -> tuple[Insumo | None, list[int]]:
    canonical = canonical_insumo_by_id(insumo_id)
    if canonical is None:
        return None, []
    for row in canonicalized_active_insumos(limit=5000):
        if canonical.id in row["member_ids"]:
            return canonical, list(row["member_ids"])
    return canonical, [canonical.id]


def _latest_cost_for_canonical(insumo_id: int | str | None, proveedor: Proveedor | None = None) -> tuple[Insumo | None, CostoInsumo | None]:
    canonical, member_ids = _canonical_member_ids(insumo_id)
    if canonical is None:
        return None, None
    costo_qs = CostoInsumo.objects.filter(insumo_id__in=member_ids).order_by("-fecha", "-id")
    if proveedor is not None:
        preferred = costo_qs.filter(proveedor=proveedor).first()
        if preferred is not None:
            return canonical, preferred
    return canonical, costo_qs.first()


def _parse_bool(value: str | None, default: bool = False) -> bool:
    if value is None:
        return default
    raw = str(value).strip().lower()
    if raw in {"1", "true", "yes", "si", "on"}:
        return True
    if raw in {"0", "false", "no", "off"}:
        return False
    return default


def _parse_period(period_raw: str | None) -> tuple[int, int] | None:
    if not period_raw:
        return None
    raw = str(period_raw).strip()
    parts = raw.split("-")
    if len(parts) != 2:
        return None
    try:
        year = int(parts[0])
        month = int(parts[1])
    except ValueError:
        return None
    if year < 2000 or year > 2200 or month < 1 or month > 12:
        return None
    return year, month


def _parse_iso_date(value: str | None) -> date | None:
    if not value:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError:
        return None


def _parse_bounded_int(raw_value, *, default: int, min_value: int, max_value: int) -> int:
    try:
        value = int(raw_value)
    except (TypeError, ValueError):
        value = default
    return max(min_value, min(value, max_value))


def _to_float(value, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError, InvalidOperation):
        return default


def _pct_change(current: int, previous: int) -> float:
    current_i = int(current or 0)
    previous_i = int(previous or 0)
    if previous_i <= 0:
        return 100.0 if current_i > 0 else 0.0
    return round(((current_i - previous_i) * 100.0 / previous_i), 2)


def _build_public_api_daily_trend(days: int = 7) -> list[dict[str, Any]]:
    days = max(1, min(int(days or 7), 31))
    today = timezone.localdate()
    start_date = today - timedelta(days=days - 1)
    raw_rows = list(
        PublicApiAccessLog.objects.filter(created_at__date__gte=start_date)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(
            total=Count("id"),
            errors=Count("id", filter=Q(status_code__gte=400)),
        )
        .order_by("day")
    )
    by_day = {row["day"]: row for row in raw_rows}
    trend = []
    for index in range(days):
        day = start_date + timedelta(days=index)
        row = by_day.get(day, {})
        total = int(row.get("total") or 0)
        errors = int(row.get("errors") or 0)
        trend.append(
            {
                "day": day,
                "requests": total,
                "errors": errors,
                "error_rate_pct": round((errors * 100.0 / total), 2) if total else 0.0,
            }
        )
    return trend


def _preview_deactivate_idle_api_clients(idle_days: int, limit: int) -> dict[str, Any]:
    idle_days = max(1, min(int(idle_days or 30), 365))
    limit = max(1, min(int(limit or 100), 500))
    cutoff = timezone.now() - timedelta(days=idle_days)
    recent_client_ids = set(
        PublicApiAccessLog.objects.filter(created_at__gte=cutoff)
        .values_list("client_id", flat=True)
        .distinct()
    )
    candidates = list(
        PublicApiClient.objects.filter(activo=True)
        .exclude(id__in=recent_client_ids)
        .order_by("id")
        .values("id", "nombre")[:limit]
    )
    return {
        "idle_days": idle_days,
        "limit": limit,
        "candidates": len(candidates),
        "deactivated": 0,
        "cutoff": cutoff.isoformat(),
        "candidate_clients": candidates,
        "dry_run": True,
    }


def _preview_purge_api_logs(retain_days: int, max_delete: int) -> dict[str, Any]:
    retain_days = max(1, min(int(retain_days or 90), 3650))
    max_delete = max(1, min(int(max_delete or 5000), 50000))
    cutoff = timezone.now() - timedelta(days=retain_days)
    total_candidates = PublicApiAccessLog.objects.filter(created_at__lt=cutoff).count()
    return {
        "retain_days": retain_days,
        "max_delete": max_delete,
        "cutoff": cutoff.isoformat(),
        "candidates": int(total_candidates),
        "deleted": 0,
        "remaining_candidates": int(total_candidates),
        "would_delete": min(int(total_candidates), max_delete),
        "dry_run": True,
    }



def _can_approve_ajustes(user) -> bool:
    return has_any_role(user, ROLE_ADMIN, ROLE_DG)


def _serialize_ajuste_row(ajuste: AjusteInventario) -> dict:
    canonical = canonical_insumo(ajuste.insumo) if ajuste.insumo_id else None
    display_insumo = canonical or ajuste.insumo
    return {
        "id": ajuste.id,
        "folio": ajuste.folio,
        "insumo_id": display_insumo.id if display_insumo else ajuste.insumo_id,
        "insumo": display_insumo.nombre if display_insumo else "",
        "insumo_canonical": bool(canonical and canonical.id != ajuste.insumo_id),
        "cantidad_sistema": str(ajuste.cantidad_sistema),
        "cantidad_fisica": str(ajuste.cantidad_fisica),
        "delta": str(_to_decimal(ajuste.cantidad_fisica) - _to_decimal(ajuste.cantidad_sistema)),
        "motivo": ajuste.motivo,
        "estatus": ajuste.estatus,
        "solicitado_por": ajuste.solicitado_por.username if ajuste.solicitado_por_id else "",
        "aprobado_por": ajuste.aprobado_por.username if ajuste.aprobado_por_id else "",
        "comentario_revision": ajuste.comentario_revision or "",
        "creado_en": ajuste.creado_en,
        "aprobado_en": ajuste.aprobado_en,
        "aplicado_en": ajuste.aplicado_en,
    }



class InventarioSugerenciasCompraView(APIView):
    permission_classes = [IsAuthenticated]

    def _resolve_scope(self, request):
        plan_id_raw = (request.GET.get("plan_id") or "").strip()
        periodo_raw = (request.GET.get("periodo") or "").strip()
        plan = None
        year = None
        month = None

        if plan_id_raw:
            try:
                plan_id = int(plan_id_raw)
            except ValueError:
                plan_id = None
            if plan_id:
                plan = get_object_or_404(PlanProduccion, pk=plan_id)
                year = plan.fecha_produccion.year
                month = plan.fecha_produccion.month

        if not plan:
            parsed = _parse_period(periodo_raw)
            if parsed:
                year, month = parsed
            else:
                today = timezone.localdate()
                year, month = today.year, today.month

        if plan:
            items_qs = PlanProduccionItem.objects.filter(plan_id=plan.id).select_related("receta", "plan")
        else:
            items_qs = PlanProduccionItem.objects.filter(
                plan__fecha_produccion__year=year,
                plan__fecha_produccion__month=month,
            ).select_related("receta", "plan")

        return items_qs, {
            "plan_id": plan.id if plan else None,
            "plan_nombre": plan.nombre if plan else "",
            "periodo": f"{year:04d}-{month:02d}",
        }

    def _requerimientos_plan(self, plan_items_qs) -> dict[int, Decimal]:
        requerimientos: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
        receta_ids = set(plan_items_qs.values_list("receta_id", flat=True))
        if not receta_ids:
            return requerimientos

        lineas_by_receta: dict[int, list[LineaReceta]] = defaultdict(list)
        lineas_qs = (
            LineaReceta.objects.filter(receta_id__in=receta_ids)
            .exclude(tipo_linea=LineaReceta.TIPO_SUBSECCION)
            .only("receta_id", "insumo_id", "cantidad")
        )
        for linea in lineas_qs:
            if not linea.insumo_id:
                continue
            qty = _to_decimal(linea.cantidad)
            if qty <= 0:
                continue
            lineas_by_receta[linea.receta_id].append(linea)

        for item in plan_items_qs:
            factor = _to_decimal(item.cantidad)
            if factor <= 0:
                continue
            for linea in lineas_by_receta.get(item.receta_id, []):
                canonical = canonical_insumo_by_id(linea.insumo_id)
                if not canonical:
                    continue
                requerimientos[canonical.id] += _to_decimal(linea.cantidad) * factor
        return requerimientos

    def _en_transito_por_insumo(self) -> dict[int, Decimal]:
        totals: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
        transit_status = [
            OrdenCompra.STATUS_ENVIADA,
            OrdenCompra.STATUS_CONFIRMADA,
            OrdenCompra.STATUS_PARCIAL,
        ]
        ordenes = (
            OrdenCompra.objects.filter(estatus__in=transit_status, solicitud__isnull=False)
            .select_related("solicitud__insumo")
            .only("solicitud__insumo_id", "solicitud__cantidad")
        )
        for orden in ordenes:
            solicitud = orden.solicitud
            if not solicitud or not solicitud.insumo_id:
                continue
            qty = _to_decimal(solicitud.cantidad)
            if qty > 0:
                canonical = canonical_insumo_by_id(solicitud.insumo_id)
                if not canonical:
                    continue
                totals[canonical.id] += qty
        return totals

    def get(self, request):
        include_all = _parse_bool(request.GET.get("include_all"), default=False)
        try:
            limit = int(request.GET.get("limit", 300))
        except ValueError:
            limit = 300
        limit = max(1, min(limit, 2000))

        plan_items_qs, scope = self._resolve_scope(request)
        requerimientos = self._requerimientos_plan(plan_items_qs)
        en_transito = self._en_transito_por_insumo()

        existencias_qs = ExistenciaInsumo.objects.select_related("insumo__unidad_base", "insumo__proveedor_principal").filter(
            insumo__activo=True
        )
        existencias_map: dict[int, dict[str, Any]] = {}
        for ex in existencias_qs:
            canonical = canonical_insumo(ex.insumo)
            if not canonical:
                continue
            row = existencias_map.setdefault(
                canonical.id,
                {
                    "canonical": canonical,
                    "stock_actual": Decimal("0"),
                    "stock_seguridad": None,
                    "punto_reorden": None,
                    "consumo_diario": None,
                    "lead_time": None,
                },
            )
            row["stock_actual"] += _to_decimal(ex.stock_actual or 0)
            if row["stock_seguridad"] is None:
                row["stock_seguridad"] = _to_decimal(ex.stock_minimo or 0)
            if row["punto_reorden"] is None:
                row["punto_reorden"] = _to_decimal(ex.punto_reorden or 0)
            if row["consumo_diario"] is None:
                row["consumo_diario"] = _to_decimal(ex.consumo_diario_promedio or 0)
            if row["lead_time"] is None:
                row["lead_time"] = int(ex.dias_llegada_pedido or 0)

        insumo_ids = set(existencias_map.keys()) | set(requerimientos.keys()) | set(en_transito.keys())
        if not insumo_ids:
            return Response(
                {
                    "scope": scope,
                    "formula": "compra_sugerida = (requerido + stock_seguridad) - (disponible + en_transito)",
                    "totales": {
                        "insumos": 0,
                        "criticos": 0,
                        "bajo_reorden": 0,
                        "compra_sugerida_total": "0",
                        "costo_estimado_total": "0",
                    },
                    "items": [],
                },
                status=status.HTTP_200_OK,
            )

        insumos = {}
        for i in Insumo.objects.filter(id__in=insumo_ids, activo=True).select_related("unidad_base", "proveedor_principal"):
            canonical = canonical_insumo(i)
            if canonical:
                insumos[canonical.id] = canonical
        if not insumos:
            return Response(
                {
                    "scope": scope,
                    "formula": "compra_sugerida = (requerido + stock_seguridad) - (disponible + en_transito)",
                    "totales": {
                        "insumos": 0,
                        "criticos": 0,
                        "bajo_reorden": 0,
                        "compra_sugerida_total": "0",
                        "costo_estimado_total": "0",
                    },
                    "items": [],
                },
                status=status.HTTP_200_OK,
            )

        latest_cost: dict[int, Decimal] = {}
        for canonical_id in insumos.keys():
            latest = latest_costo_canonico(insumo_id=canonical_id)
            if latest is not None:
                latest_cost[canonical_id] = _to_decimal(latest)

        rows = []
        total_sugerido = Decimal("0")
        total_costo = Decimal("0")
        criticos = 0
        bajo_reorden = 0

        for insumo_id in sorted(insumos.keys(), key=lambda pk: insumos[pk].nombre.lower()):
            insumo = insumos[insumo_id]
            ex = existencias_map.get(insumo_id)

            stock_actual = _to_decimal(ex["stock_actual"] if ex else 0)
            stock_seguridad = _to_decimal(ex["stock_seguridad"] if ex and ex["stock_seguridad"] is not None else 0)
            punto_reorden = _to_decimal(ex["punto_reorden"] if ex and ex["punto_reorden"] is not None else 0)
            consumo_diario = _to_decimal(ex["consumo_diario"] if ex and ex["consumo_diario"] is not None else 0)
            lead_time = int(ex["lead_time"] or 0) if ex else 0
            if lead_time <= 0 and insumo.proveedor_principal_id:
                lead_time = int(insumo.proveedor_principal.lead_time_dias or 0)
            lead_time = max(lead_time, 0)

            demanda_lead_time = consumo_diario * Decimal(str(lead_time))
            requerido_plan = _to_decimal(requerimientos.get(insumo_id, 0))
            requerido = requerido_plan if requerido_plan > demanda_lead_time else demanda_lead_time
            en_transito_qty = _to_decimal(en_transito.get(insumo_id, 0))

            sugerida = requerido + stock_seguridad - (stock_actual + en_transito_qty)
            if sugerida < 0:
                sugerida = Decimal("0")

            costo_unitario = _to_decimal(latest_cost.get(insumo_id, 0))
            costo_sugerido = sugerida * costo_unitario if sugerida > 0 and costo_unitario > 0 else Decimal("0")

            if stock_actual <= 0 and (requerido > 0 or punto_reorden > 0):
                estado = "CRITICO"
                criticos += 1
            elif punto_reorden > 0 and stock_actual < punto_reorden:
                estado = "BAJO_REORDEN"
                bajo_reorden += 1
            else:
                estado = "SUFICIENTE"

            if not include_all and sugerida <= 0:
                continue

            total_sugerido += sugerida
            total_costo += costo_sugerido

            rows.append(
                {
                    "insumo_id": insumo_id,
                    "insumo": insumo.nombre,
                    "unidad": insumo.unidad_base.codigo if insumo.unidad_base_id and insumo.unidad_base else "",
                    "proveedor_principal": insumo.proveedor_principal.nombre if insumo.proveedor_principal_id else "",
                    "stock_actual": str(stock_actual),
                    "stock_seguridad": str(stock_seguridad),
                    "punto_reorden": str(punto_reorden),
                    "en_transito": str(en_transito_qty),
                    "requerido_plan": str(requerido_plan),
                    "demanda_lead_time": str(demanda_lead_time),
                    "requerido_total": str(requerido),
                    "compra_sugerida": str(sugerida),
                    "lead_time_dias": lead_time,
                    "consumo_diario_promedio": str(consumo_diario),
                    "costo_unitario": str(costo_unitario),
                    "costo_compra_sugerida": str(costo_sugerido),
                    "estatus": estado,
                }
            )

        rows.sort(
            key=lambda x: (
                Decimal(x["compra_sugerida"]),
                Decimal(x["requerido_total"]),
                x["insumo"].lower(),
            ),
            reverse=True,
        )

        return Response(
            {
                "scope": {**scope, "include_all": include_all, "limit": limit},
                "formula": "compra_sugerida = (requerido + stock_seguridad) - (disponible + en_transito)",
                "totales": {
                    "insumos": len(rows[:limit]),
                    "criticos": criticos,
                    "bajo_reorden": bajo_reorden,
                    "compra_sugerida_total": str(total_sugerido),
                    "costo_estimado_total": str(total_costo),
                },
                "items": rows[:limit],
            },
            status=status.HTTP_200_OK,
        )


class InventarioAliasesListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _serialize_alias(alias: InsumoAlias) -> dict:
        insumo = canonical_insumo(alias.insumo) if alias.insumo_id else None
        return {
            "id": alias.id,
            "alias": alias.nombre,
            "normalizado": alias.nombre_normalizado,
            "insumo_id": insumo.id if insumo else alias.insumo_id,
            "insumo": insumo.nombre if insumo else "",
            "unidad": insumo.unidad_base.codigo if insumo and insumo.unidad_base_id else "",
            "categoria": insumo.categoria if insumo else "",
        }

    def get(self, request):
        if not can_view_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar aliases de inventario."},
                status=status.HTTP_403_FORBIDDEN,
            )

        q = (request.GET.get("q") or "").strip()
        insumo_id_raw = (request.GET.get("insumo_id") or "").strip()
        limit = _parse_bounded_int(request.GET.get("limit", 250), default=250, min_value=1, max_value=1200)

        qs = InsumoAlias.objects.select_related("insumo__unidad_base").order_by("nombre")
        if q:
            q_norm = normalizar_nombre(q)
            qs = qs.filter(
                Q(nombre__icontains=q)
                | Q(nombre_normalizado__icontains=q_norm)
                | Q(insumo__nombre__icontains=q)
                | Q(insumo__codigo_point__icontains=q)
            )

        insumo_id = None
        if insumo_id_raw:
            try:
                insumo_id = int(insumo_id_raw)
            except ValueError:
                return Response(
                    {"detail": "insumo_id inválido."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            insumo_target = canonical_insumo_by_id(insumo_id)
            if not insumo_target:
                return Response(
                    {"detail": "insumo_id no encontrado o inactivo."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            insumo_id = insumo_target.id
            qs = qs.filter(insumo_id=insumo_id)

        total_rows = qs.count()
        rows = [self._serialize_alias(a) for a in qs[:limit]]
        return Response(
            {
                "filters": {"q": q, "insumo_id": insumo_id, "limit": limit},
                "totales": {"rows": total_rows, "returned": len(rows)},
                "items": rows,
            },
            status=status.HTTP_200_OK,
        )

    def post(self, request):
        if not can_manage_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para crear/editar aliases de inventario."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = InventarioAliasCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        alias_name = (data.get("alias") or "").strip()
        alias_norm = normalizar_nombre(alias_name)
        if not alias_norm:
            return Response(
                {"detail": "Alias inválido: nombre vacío después de normalizar."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        insumo = canonical_insumo_by_id(data["insumo_id"])
        if insumo is None:
            return Response(
                {"detail": "insumo_id no encontrado o inactivo."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        alias, created = InsumoAlias.objects.get_or_create(
            nombre_normalizado=alias_norm,
            defaults={"nombre": alias_name[:250], "insumo": insumo},
        )

        updated = False
        if not created:
            changed = []
            if alias.insumo_id != insumo.id:
                alias.insumo = insumo
                changed.append("insumo")
            if alias.nombre != alias_name[:250]:
                alias.nombre = alias_name[:250]
                changed.append("nombre")
            if changed:
                alias.save(update_fields=changed)
                updated = True

        point_resolved = 0
        recetas_resolved = 0
        if bool(data.get("resolver_cross_source", True)):
            point_resolved, recetas_resolved = _resolve_cross_source_with_alias(alias_name, insumo)

        alias.refresh_from_db()
        return Response(
            {
                "created": created,
                "updated": updated,
                "alias": self._serialize_alias(alias),
                "resolved": {
                    "point_pending": point_resolved,
                    "recetas_pending": recetas_resolved,
                },
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )


class InventarioAliasesMassReassignView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not can_manage_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para reasignar aliases de inventario."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = InventarioAliasMassReassignSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        insumo = canonical_insumo_by_id(data["insumo_id"])
        if insumo is None:
            return Response(
                {"detail": "insumo_id no encontrado o inactivo."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        alias_ids = list(dict.fromkeys(int(aid) for aid in data["alias_ids"]))
        aliases = list(InsumoAlias.objects.filter(id__in=alias_ids).select_related("insumo"))
        if not aliases:
            return Response(
                {"detail": "No se encontraron aliases con los IDs enviados."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        resolver_cross = bool(data.get("resolver_cross_source", True))
        updated = 0
        point_resolved = 0
        recetas_resolved = 0
        touched_ids: list[int] = []
        for alias in aliases:
            if alias.insumo_id == insumo.id:
                continue
            alias.insumo = insumo
            alias.save(update_fields=["insumo"])
            updated += 1
            touched_ids.append(alias.id)
            if resolver_cross:
                p_count, r_count = _resolve_cross_source_with_alias(alias.nombre, insumo)
                point_resolved += p_count
                recetas_resolved += r_count

        return Response(
            {
                "target_insumo": {"id": insumo.id, "nombre": insumo.nombre},
                "selected": len(alias_ids),
                "found": len(aliases),
                "updated": updated,
                "touched_alias_ids": touched_ids,
                "resolved": {
                    "point_pending": point_resolved,
                    "recetas_pending": recetas_resolved,
                },
            },
            status=status.HTTP_200_OK,
        )


class InventarioAliasesPendientesView(APIView):
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _fmt_datetime(value):
        if not value:
            return ""
        try:
            if timezone.is_aware(value):
                value = timezone.localtime(value)
            return value.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(value)

    def _export_csv(self, almacen_rows: list[dict], point_rows: list[dict], recetas_rows: list[dict]) -> HttpResponse:
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="inventario_aliases_pendientes.csv"'
        writer = csv.writer(response)
        writer.writerow(
            [
                "seccion",
                "id",
                "tipo",
                "codigo",
                "nombre",
                "nombre_normalizado",
                "sugerencia",
                "score",
                "metodo",
                "extra",
            ]
        )
        for row in almacen_rows:
            writer.writerow(
                [
                    "almacen",
                    row.get("run_id", ""),
                    "",
                    "",
                    row.get("nombre_origen", ""),
                    row.get("nombre_normalizado", ""),
                    row.get("sugerencia", ""),
                    row.get("score", ""),
                    row.get("metodo", ""),
                    row.get("fuente", ""),
                ]
            )
        for row in point_rows:
            writer.writerow(
                [
                    "point",
                    row.get("id", ""),
                    row.get("tipo", ""),
                    row.get("point_codigo", ""),
                    row.get("point_nombre", ""),
                    "",
                    row.get("sugerencia", ""),
                    row.get("score", ""),
                    row.get("metodo", ""),
                    "",
                ]
            )
        for row in recetas_rows:
            writer.writerow(
                [
                    "recetas",
                    row.get("id", ""),
                    row.get("estatus", ""),
                    "",
                    row.get("insumo_texto", ""),
                    row.get("nombre_normalizado", ""),
                    "",
                    row.get("score", ""),
                    row.get("metodo", ""),
                    row.get("receta", ""),
                ]
            )
        return response

    def _export_xlsx(self, almacen_rows: list[dict], point_rows: list[dict], recetas_rows: list[dict]) -> HttpResponse:
        wb = Workbook()
        ws_alm = wb.active
        ws_alm.title = "almacen"
        ws_alm.append(
            [
                "run_id",
                "run_started_at",
                "nombre_origen",
                "nombre_normalizado",
                "sugerencia",
                "score",
                "metodo",
                "fuente",
            ]
        )
        for row in almacen_rows:
            ws_alm.append(
                [
                    row.get("run_id", ""),
                    self._fmt_datetime(row.get("run_started_at")),
                    row.get("nombre_origen", ""),
                    row.get("nombre_normalizado", ""),
                    row.get("sugerencia", ""),
                    row.get("score", 0),
                    row.get("metodo", ""),
                    row.get("fuente", ""),
                ]
            )

        ws_point = wb.create_sheet("point")
        ws_point.append(["id", "tipo", "point_codigo", "point_nombre", "sugerencia", "score", "metodo", "actualizado_en"])
        for row in point_rows:
            ws_point.append(
                [
                    row.get("id", ""),
                    row.get("tipo", ""),
                    row.get("point_codigo", ""),
                    row.get("point_nombre", ""),
                    row.get("sugerencia", ""),
                    row.get("score", 0),
                    row.get("metodo", ""),
                    self._fmt_datetime(row.get("actualizado_en")),
                ]
            )

        ws_rec = wb.create_sheet("recetas")
        ws_rec.append(["id", "receta_id", "receta", "insumo_texto", "nombre_normalizado", "score", "metodo", "estatus"])
        for row in recetas_rows:
            ws_rec.append(
                [
                    row.get("id", ""),
                    row.get("receta_id", ""),
                    row.get("receta", ""),
                    row.get("insumo_texto", ""),
                    row.get("nombre_normalizado", ""),
                    row.get("score", 0),
                    row.get("metodo", ""),
                    row.get("estatus", ""),
                ]
            )

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="inventario_aliases_pendientes.xlsx"'
        return response

    def get(self, request):
        if not can_view_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar pendientes de homologación."},
                status=status.HTTP_403_FORBIDDEN,
            )

        limit = _parse_bounded_int(request.GET.get("limit", 120), default=120, min_value=1, max_value=400)
        offset = _parse_bounded_int(request.GET.get("offset", 0), default=0, min_value=0, max_value=50000)
        runs_to_scan = _parse_bounded_int(request.GET.get("runs", 5), default=5, min_value=1, max_value=30)
        runs_detail = _parse_bounded_int(request.GET.get("runs_detail", 5), default=5, min_value=1, max_value=20)
        include_runs = _parse_bool(request.GET.get("include_runs"), default=True)
        q = (request.GET.get("q") or "").strip()
        q_norm = normalizar_nombre(q)
        source = (request.GET.get("source") or "TODOS").strip().upper()
        export = (request.GET.get("export") or "").strip().lower()
        point_tipo = (request.GET.get("point_tipo") or PointPendingMatch.TIPO_INSUMO).strip().upper()
        valid_point_tipos = {
            PointPendingMatch.TIPO_INSUMO,
            PointPendingMatch.TIPO_PROVEEDOR,
            PointPendingMatch.TIPO_PRODUCTO,
            "TODOS",
            "ALL",
        }
        valid_sources = {"TODOS", "ALL", "ALMACEN", "POINT", "RECETAS"}
        if export not in {"", "csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if source not in valid_sources:
            return Response(
                {"detail": "source inválido. Usa ALMACEN, POINT, RECETAS o TODOS."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if point_tipo not in valid_point_tipos:
            return Response(
                {"detail": "point_tipo inválido. Usa INSUMO, PROVEEDOR, PRODUCTO o TODOS."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        almacen_rows_all: list[dict] = []
        recent_runs: list[dict] = []
        sync_runs = list(
            AlmacenSyncRun.objects.only(
                "id",
                "started_at",
                "source",
                "status",
                "matched",
                "unmatched",
                "pending_preview",
            ).order_by("-started_at")[:runs_to_scan]
        )
        if include_runs:
            for run in sync_runs[:runs_detail]:
                recent_runs.append(
                    {
                        "id": run.id,
                        "started_at": run.started_at,
                        "source": run.source,
                        "status": run.status,
                        "matched": int(run.matched or 0),
                        "unmatched": int(run.unmatched or 0),
                        "has_preview": bool(run.pending_preview),
                    }
                )
        for run in sync_runs:
            for row in run.pending_preview or []:
                nombre_origen = str((row or {}).get("nombre_origen") or "").strip()
                if not nombre_origen:
                    continue
                almacen_rows_all.append(
                    {
                        "run_id": run.id,
                        "run_started_at": run.started_at,
                        "nombre_origen": nombre_origen,
                        "nombre_normalizado": str((row or {}).get("nombre_normalizado") or normalizar_nombre(nombre_origen)),
                        "sugerencia": str((row or {}).get("suggestion") or ""),
                        "score": float((row or {}).get("score") or 0),
                        "metodo": str((row or {}).get("method") or ""),
                        "fuente": str((row or {}).get("fuente") or "ALMACEN"),
                    }
                )
        if q_norm:
            almacen_rows_all = [
                row
                for row in almacen_rows_all
                if q_norm in normalizar_nombre(row.get("nombre_origen") or "")
                or q_norm in normalizar_nombre(row.get("nombre_normalizado") or "")
                or q_norm in normalizar_nombre(row.get("sugerencia") or "")
            ]
        almacen_total = len(almacen_rows_all)

        point_qs = PointPendingMatch.objects.order_by("-fuzzy_score", "point_nombre")
        if not q_norm:
            point_qs = point_qs.visible_en_operacion()
        if point_tipo not in {"TODOS", "ALL"}:
            point_qs = point_qs.filter(tipo=point_tipo)
        if q_norm:
            point_qs = point_qs.filter(
                Q(point_nombre__icontains=q)
                | Q(point_codigo__icontains=q)
                | Q(fuzzy_sugerencia__icontains=q)
                | Q(tipo__icontains=q)
            )
        point_total = point_qs.count()
        point_totals_by_tipo = {
            row["tipo"]: row["count"]
            for row in (
                point_qs.values("tipo")
                .annotate(count=Count("id"))
                .order_by("tipo")
            )
        }
        point_source_qs = point_qs if export else point_qs[offset : offset + limit]
        point_rows = [
            {
                "id": p.id,
                "tipo": p.tipo,
                "point_codigo": p.point_codigo,
                "point_nombre": p.point_nombre,
                "sugerencia": p.fuzzy_sugerencia or "",
                "score": float(p.fuzzy_score or 0),
                "metodo": p.method or "",
                "actualizado_en": p.actualizado_en,
            }
            for p in point_source_qs
        ]

        recetas_qs = (
            LineaReceta.objects.filter(insumo__isnull=True)
            .filter(match_status__in=[LineaReceta.STATUS_NEEDS_REVIEW, LineaReceta.STATUS_REJECTED])
            .select_related("receta")
            .order_by("-match_score", "receta__nombre", "posicion")
        )
        if q_norm:
            recetas_qs = recetas_qs.filter(
                Q(receta__nombre__icontains=q)
                | Q(insumo_texto__icontains=q)
            )
        recetas_total = recetas_qs.count()
        recetas_source_qs = recetas_qs if export else recetas_qs[offset : offset + limit]
        recetas_rows = [
            {
                "id": linea.id,
                "receta_id": linea.receta_id,
                "receta": linea.receta.nombre,
                "insumo_texto": linea.insumo_texto or "",
                "nombre_normalizado": normalizar_nombre(linea.insumo_texto or ""),
                "score": float(linea.match_score or 0),
                "metodo": linea.match_method or "",
                "estatus": linea.match_status,
            }
            for linea in recetas_source_qs
        ]
        almacen_rows = list(almacen_rows_all if export else almacen_rows_all[offset : offset + limit])
        if source == "ALMACEN":
            point_rows = []
            recetas_rows = []
            point_total = 0
            point_totals_by_tipo = {}
            recetas_total = 0
            almacen_total = len(almacen_rows_all)
        elif source == "POINT":
            almacen_rows = []
            recetas_rows = []
            almacen_total = 0
            recetas_total = 0
        elif source == "RECETAS":
            almacen_rows = []
            point_rows = []
            almacen_total = 0
            point_total = 0
            point_totals_by_tipo = {}

        if export == "csv":
            export_almacen = list(almacen_rows_all if source in {"TODOS", "ALL", "ALMACEN"} else [])
            export_point = point_rows
            export_recetas = recetas_rows
            if source in {"TODOS", "ALL"}:
                export_point = [
                    {
                        "id": p.id,
                        "tipo": p.tipo,
                        "point_codigo": p.point_codigo,
                        "point_nombre": p.point_nombre,
                        "sugerencia": p.fuzzy_sugerencia or "",
                        "score": float(p.fuzzy_score or 0),
                        "metodo": p.method or "",
                        "actualizado_en": p.actualizado_en,
                    }
                    for p in point_qs
                ]
                export_recetas = [
                    {
                        "id": linea.id,
                        "receta_id": linea.receta_id,
                        "receta": linea.receta.nombre,
                        "insumo_texto": linea.insumo_texto or "",
                        "nombre_normalizado": normalizar_nombre(linea.insumo_texto or ""),
                        "score": float(linea.match_score or 0),
                        "metodo": linea.match_method or "",
                        "estatus": linea.match_status,
                    }
                    for linea in recetas_qs
                ]
            if source == "POINT":
                export_almacen = []
                export_recetas = []
            elif source == "RECETAS":
                export_almacen = []
                export_point = []
            return self._export_csv(export_almacen, export_point, export_recetas)
        if export == "xlsx":
            export_almacen = list(almacen_rows_all if source in {"TODOS", "ALL", "ALMACEN"} else [])
            export_point = point_rows
            export_recetas = recetas_rows
            if source in {"TODOS", "ALL"}:
                export_point = [
                    {
                        "id": p.id,
                        "tipo": p.tipo,
                        "point_codigo": p.point_codigo,
                        "point_nombre": p.point_nombre,
                        "sugerencia": p.fuzzy_sugerencia or "",
                        "score": float(p.fuzzy_score or 0),
                        "metodo": p.method or "",
                        "actualizado_en": p.actualizado_en,
                    }
                    for p in point_qs
                ]
                export_recetas = [
                    {
                        "id": linea.id,
                        "receta_id": linea.receta_id,
                        "receta": linea.receta.nombre,
                        "insumo_texto": linea.insumo_texto or "",
                        "nombre_normalizado": normalizar_nombre(linea.insumo_texto or ""),
                        "score": float(linea.match_score or 0),
                        "metodo": linea.match_method or "",
                        "estatus": linea.match_status,
                    }
                    for linea in recetas_qs
                ]
            if source == "POINT":
                export_almacen = []
                export_recetas = []
            elif source == "RECETAS":
                export_almacen = []
                export_point = []
            return self._export_xlsx(export_almacen, export_point, export_recetas)

        def _section_pagination(total: int, returned: int) -> dict:
            has_next = (offset + returned) < total
            has_prev = offset > 0
            return {
                "total": total,
                "returned": returned,
                "has_next": has_next,
                "next_offset": offset + limit if has_next else None,
                "has_prev": has_prev,
                "prev_offset": max(offset - limit, 0) if has_prev else None,
            }

        return Response(
            {
                "filters": {
                    "limit": limit,
                    "offset": offset,
                    "runs": runs_to_scan,
                    "runs_detail": runs_detail,
                    "include_runs": include_runs,
                    "q": q,
                    "source": source,
                    "point_tipo": point_tipo,
                    "export": export,
                },
                "recent_runs": recent_runs,
                "totales": {
                    "almacen": almacen_total,
                    "point": point_total,
                    "point_by_tipo": point_totals_by_tipo,
                    "recetas": recetas_total,
                },
                "pagination": {
                    "limit": limit,
                    "offset": offset,
                    "almacen": _section_pagination(almacen_total, len(almacen_rows)),
                    "point": _section_pagination(point_total, len(point_rows)),
                    "recetas": _section_pagination(recetas_total, len(recetas_rows)),
                },
                "items": {
                    "almacen": almacen_rows,
                    "point": point_rows,
                    "recetas": recetas_rows,
                },
            },
            status=status.HTTP_200_OK,
        )


class InventarioAliasesPendientesUnificadosView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not can_view_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar pendientes de homologación."},
                status=status.HTTP_403_FORBIDDEN,
            )

        limit = _parse_bounded_int(request.GET.get("limit", 120), default=120, min_value=1, max_value=600)
        offset = _parse_bounded_int(request.GET.get("offset", 0), default=0, min_value=0, max_value=50000)
        runs_to_scan = _parse_bounded_int(request.GET.get("runs", 5), default=5, min_value=1, max_value=30)
        q = (request.GET.get("q") or "").strip()
        q_norm = normalizar_nombre(q)
        source = (request.GET.get("source") or "TODOS").strip().upper()
        point_tipo = (request.GET.get("point_tipo") or PointPendingMatch.TIPO_INSUMO).strip().upper()
        only_suggested = _parse_bool(request.GET.get("only_suggested"), default=False)
        min_sources = _parse_bounded_int(request.GET.get("min_sources", 1), default=1, min_value=1, max_value=3)
        score_min = float(_to_decimal(request.GET.get("score_min"), Decimal("0")))
        score_min = max(0.0, min(100.0, score_min))
        sort_by = (request.GET.get("sort_by") or "sources_active").strip().lower()
        sort_dir = (request.GET.get("sort_dir") or "desc").strip().lower()
        export = (request.GET.get("export") or "").strip().lower()
        valid_point_tipos = {
            PointPendingMatch.TIPO_INSUMO,
            PointPendingMatch.TIPO_PROVEEDOR,
            PointPendingMatch.TIPO_PRODUCTO,
            "TODOS",
            "ALL",
        }
        valid_sources = {"TODOS", "ALL", "ALMACEN", "POINT", "RECETAS"}
        if export not in {"", "csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if source not in valid_sources:
            return Response(
                {"detail": "source inválido. Usa ALMACEN, POINT, RECETAS o TODOS."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if point_tipo not in valid_point_tipos:
            return Response(
                {"detail": "point_tipo inválido. Usa INSUMO, PROVEEDOR, PRODUCTO o TODOS."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        point_tipos_filter = (
            None
            if point_tipo in {"TODOS", "ALL"}
            else [point_tipo]
        )
        allowed_sort = {
            "sources_active": lambda row: int(row.get("sources_active") or 0),
            "total_count": lambda row: int(row.get("total_count") or 0),
            "score_max": lambda row: float(row.get("score_max") or 0.0),
            "point_count": lambda row: int(row.get("point_count") or 0),
            "almacen_count": lambda row: int(row.get("almacen_count") or 0),
            "receta_count": lambda row: int(row.get("receta_count") or 0),
            "nombre_muestra": lambda row: str(row.get("nombre_muestra") or "").lower(),
            "nombre_normalizado": lambda row: str(row.get("nombre_normalizado") or "").lower(),
        }
        if sort_by not in allowed_sort:
            return Response(
                {
                    "detail": (
                        "sort_by inválido. Usa sources_active, total_count, score_max, point_count, "
                        "almacen_count, receta_count, nombre_muestra o nombre_normalizado."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        if sort_dir not in {"asc", "desc"}:
            return Response(
                {"detail": "sort_dir inválido. Usa asc o desc."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        pending_rows: list[dict] = []
        sync_runs = list(AlmacenSyncRun.objects.only("id", "started_at", "pending_preview").order_by("-started_at")[:runs_to_scan])
        for run in sync_runs:
            for row in run.pending_preview or []:
                nombre_origen = str((row or {}).get("nombre_origen") or "").strip()
                if not nombre_origen:
                    continue
                pending_rows.append(
                    {
                        "nombre_origen": nombre_origen,
                        "nombre_normalizado": str((row or {}).get("nombre_normalizado") or normalizar_nombre(nombre_origen)),
                        "sugerencia": str((row or {}).get("suggestion") or ""),
                        "score": float((row or {}).get("score") or 0),
                        "source": str((row or {}).get("fuente") or "ALMACEN"),
                    }
                )

        pending_grouped = _build_pending_grouped(pending_rows)
        unified_rows, point_unmatched_count, receta_pending_lines = _build_cross_unified_rows(
            pending_grouped,
            point_tipos=point_tipos_filter,
        )
        filtered_rows = _apply_cross_filters(
            unified_rows,
            cross_q_norm=q_norm,
            cross_only_suggested=only_suggested,
            cross_min_sources=min_sources,
            cross_score_min=score_min,
        )
        if source == "ALMACEN":
            filtered_rows = [row for row in filtered_rows if int(row.get("almacen_count") or 0) > 0]
        elif source == "POINT":
            filtered_rows = [row for row in filtered_rows if int(row.get("point_count") or 0) > 0]
        elif source == "RECETAS":
            filtered_rows = [row for row in filtered_rows if int(row.get("receta_count") or 0) > 0]

        sort_key = allowed_sort[sort_by]
        reverse = sort_dir == "desc"
        filtered_rows = sorted(
            filtered_rows,
            key=lambda row: (sort_key(row), str(row.get("nombre_muestra") or "").lower()),
            reverse=reverse,
        )
        if export == "csv":
            return _export_cross_pending_csv(filtered_rows)
        if export == "xlsx":
            return _export_cross_pending_xlsx(filtered_rows)

        overlap_2_plus = sum(1 for row in unified_rows if int(row.get("sources_active") or 0) >= 2)
        items = filtered_rows[offset : offset + limit]
        has_next = (offset + len(items)) < len(filtered_rows)
        has_prev = offset > 0
        next_offset = offset + limit if has_next else None
        prev_offset = max(offset - limit, 0) if has_prev else None

        return Response(
            {
                "filters": {
                    "limit": limit,
                    "offset": offset,
                    "runs": runs_to_scan,
                    "q": q,
                    "source": source,
                    "point_tipo": point_tipo,
                    "min_sources": min_sources,
                    "score_min": round(score_min, 2),
                    "only_suggested": only_suggested,
                    "sort_by": sort_by,
                    "sort_dir": sort_dir,
                    "export": export,
                },
                "totales": {
                    "runs_scanned": len(sync_runs),
                    "almacen_rows_raw": len(pending_rows),
                    "almacen_grouped": len(pending_grouped),
                    "unified_total": len(unified_rows),
                    "filtered_total": len(filtered_rows),
                    "returned": len(items),
                    "overlap_2_plus": overlap_2_plus,
                    "point_unmatched": point_unmatched_count,
                    "recetas_pending_lines": receta_pending_lines,
                },
                "pagination": {
                    "limit": limit,
                    "offset": offset,
                    "has_next": has_next,
                    "next_offset": next_offset,
                    "has_prev": has_prev,
                    "prev_offset": prev_offset,
                },
                "items": items,
            },
            status=status.HTTP_200_OK,
        )


class InventarioAliasesPendientesUnificadosResolveView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not has_any_role(request.user, ROLE_ADMIN, ROLE_DG, ROLE_COMPRAS):
            return Response(
                {"detail": "No tienes permisos para resolver pendientes unificados."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = InventarioCrossPendientesResolveSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        limit = int(data.get("limit") or 300)
        offset = int(data.get("offset") or 0)
        runs_to_scan = int(data.get("runs") or 5)
        q = str(data.get("q") or "").strip()
        q_norm = normalizar_nombre(q)
        source = str(data.get("source") or "TODOS").strip().upper()
        point_tipo = str(data.get("point_tipo") or PointPendingMatch.TIPO_INSUMO).strip().upper()
        sort_by = str(data.get("sort_by") or "sources_active").strip().lower()
        sort_dir = str(data.get("sort_dir") or "desc").strip().lower()
        min_sources = int(data.get("min_sources") or 2)
        score_min = float(data.get("score_min") or 0)
        score_min = max(0.0, min(100.0, score_min))
        only_suggested = bool(data.get("only_suggested", True))
        dry_run = bool(data.get("dry_run", True))
        nombres = [str(x or "").strip() for x in (data.get("nombres") or [])]
        selected_norms = {normalizar_nombre(x) for x in nombres if normalizar_nombre(x)}
        point_tipos_filter = (
            None
            if point_tipo in {"TODOS", "ALL"}
            else [point_tipo]
        )

        pending_rows: list[dict] = []
        sync_runs = list(AlmacenSyncRun.objects.only("id", "started_at", "pending_preview").order_by("-started_at")[:runs_to_scan])
        for run in sync_runs:
            for row in run.pending_preview or []:
                nombre_origen = str((row or {}).get("nombre_origen") or "").strip()
                if not nombre_origen:
                    continue
                pending_rows.append(
                    {
                        "nombre_origen": nombre_origen,
                        "nombre_normalizado": str((row or {}).get("nombre_normalizado") or normalizar_nombre(nombre_origen)),
                        "sugerencia": str((row or {}).get("suggestion") or ""),
                        "score": float((row or {}).get("score") or 0),
                        "source": str((row or {}).get("fuente") or "ALMACEN"),
                    }
                )

        pending_grouped = _build_pending_grouped(pending_rows)
        unified_rows, _, _ = _build_cross_unified_rows(
            pending_grouped,
            point_tipos=point_tipos_filter,
        )
        filtered_rows = _apply_cross_filters(
            unified_rows,
            cross_q_norm=q_norm,
            cross_only_suggested=only_suggested,
            cross_min_sources=min_sources,
            cross_score_min=score_min,
        )
        if source == "ALMACEN":
            filtered_rows = [row for row in filtered_rows if int(row.get("almacen_count") or 0) > 0]
        elif source == "POINT":
            filtered_rows = [row for row in filtered_rows if int(row.get("point_count") or 0) > 0]
        elif source == "RECETAS":
            filtered_rows = [row for row in filtered_rows if int(row.get("receta_count") or 0) > 0]

        if selected_norms:
            filtered_rows = [row for row in filtered_rows if (row.get("nombre_normalizado") or "") in selected_norms]
        allowed_sort = {
            "sources_active": lambda row: int(row.get("sources_active") or 0),
            "total_count": lambda row: int(row.get("total_count") or 0),
            "score_max": lambda row: float(row.get("score_max") or 0.0),
            "point_count": lambda row: int(row.get("point_count") or 0),
            "almacen_count": lambda row: int(row.get("almacen_count") or 0),
            "receta_count": lambda row: int(row.get("receta_count") or 0),
            "nombre_muestra": lambda row: str(row.get("nombre_muestra") or "").lower(),
            "nombre_normalizado": lambda row: str(row.get("nombre_normalizado") or "").lower(),
        }
        sort_key = allowed_sort[sort_by]
        reverse = sort_dir == "desc"
        filtered_rows = sorted(
            filtered_rows,
            key=lambda row: (sort_key(row), str(row.get("nombre_muestra") or "").lower()),
            reverse=reverse,
        )
        rows_to_process = filtered_rows[offset : offset + limit]

        processed = 0
        resolved = 0
        created_aliases = 0
        updated_aliases = 0
        preview_create_aliases = 0
        preview_update_aliases = 0
        unchanged = 0
        skipped_no_suggestion = 0
        skipped_no_target = 0
        point_resolved_total = 0
        recetas_resolved_total = 0
        preview_actions: list[dict] = []

        write_context = nullcontext()
        if not dry_run:
            write_context = transaction.atomic()

        with write_context:
            for row in rows_to_process:
                processed += 1
                alias_name = str(row.get("nombre_muestra") or "").strip()
                alias_norm = normalizar_nombre(alias_name)
                suggestion_name = str(row.get("suggestion") or "").strip()
                suggestion_norm = normalizar_nombre(suggestion_name)
                if not suggestion_norm:
                    skipped_no_suggestion += 1
                    if len(preview_actions) < 120:
                        preview_actions.append(
                            {
                                "nombre_muestra": alias_name,
                                "sugerencia": suggestion_name,
                                "action": "skip_no_suggestion",
                            }
                        )
                    continue

                insumo_target = Insumo.objects.filter(activo=True, nombre_normalizado=suggestion_norm).first()
                if not insumo_target:
                    skipped_no_target += 1
                    if len(preview_actions) < 120:
                        preview_actions.append(
                            {
                                "nombre_muestra": alias_name,
                                "sugerencia": suggestion_name,
                                "action": "skip_no_target",
                            }
                        )
                    continue

                action_name = "noop"
                if not alias_norm or alias_norm == insumo_target.nombre_normalizado:
                    unchanged += 1
                    action_name = "noop_same_name"
                else:
                    alias_obj = InsumoAlias.objects.filter(nombre_normalizado=alias_norm).first()
                    if dry_run:
                        if alias_obj is None:
                            action_name = "create_alias"
                            preview_create_aliases += 1
                        elif alias_obj.insumo_id != insumo_target.id or alias_obj.nombre != alias_name[:250]:
                            action_name = "update_alias"
                            preview_update_aliases += 1
                        else:
                            action_name = "noop_alias_exists"
                            unchanged += 1
                    else:
                        if alias_obj is None:
                            InsumoAlias.objects.create(
                                nombre=alias_name[:250],
                                nombre_normalizado=alias_norm,
                                insumo=insumo_target,
                            )
                            created_aliases += 1
                            action_name = "create_alias"
                        else:
                            changes = []
                            if alias_obj.insumo_id != insumo_target.id:
                                alias_obj.insumo = insumo_target
                                changes.append("insumo")
                            if alias_obj.nombre != alias_name[:250]:
                                alias_obj.nombre = alias_name[:250]
                                changes.append("nombre")
                            if changes:
                                alias_obj.save(update_fields=changes)
                                updated_aliases += 1
                                action_name = "update_alias"
                            else:
                                unchanged += 1
                                action_name = "noop_alias_exists"

                if not dry_run:
                    p_count, r_count = _resolve_cross_source_with_alias(alias_name or suggestion_name, insumo_target)
                    point_resolved_total += p_count
                    recetas_resolved_total += r_count

                resolved += 1
                if len(preview_actions) < 120:
                    preview_actions.append(
                        {
                            "nombre_muestra": alias_name,
                            "sugerencia": suggestion_name,
                            "insumo_target": insumo_target.nombre,
                            "insumo_id": insumo_target.id,
                            "action": action_name,
                            "sources_active": int(row.get("sources_active") or 0),
                            "total_count": int(row.get("total_count") or 0),
                        }
                    )

        return Response(
            {
                "dry_run": dry_run,
                "filters": {
                    "q": q,
                    "runs": runs_to_scan,
                    "limit": limit,
                    "offset": offset,
                    "source": source,
                    "point_tipo": point_tipo,
                    "min_sources": min_sources,
                    "score_min": round(score_min, 2),
                    "only_suggested": only_suggested,
                    "sort_by": sort_by,
                    "sort_dir": sort_dir,
                    "selected_names_count": len(selected_norms),
                },
                "totales": {
                    "candidatos_filtrados": len(filtered_rows),
                    "candidatos_pagina": len(rows_to_process),
                    "procesados": processed,
                    "resueltos": resolved,
                    "aliases_creados": created_aliases,
                    "aliases_actualizados": updated_aliases,
                    "aliases_creados_preview": (preview_create_aliases if dry_run else created_aliases),
                    "aliases_actualizados_preview": (preview_update_aliases if dry_run else updated_aliases),
                    "sin_cambio": unchanged,
                    "sin_sugerencia": skipped_no_suggestion,
                    "sin_insumo_objetivo": skipped_no_target,
                    "point_resueltos": point_resolved_total,
                    "recetas_resueltas": recetas_resolved_total,
                },
                "items": preview_actions,
            },
            status=status.HTTP_200_OK,
        )



class InventarioPointPendingResolveView(APIView):
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _resolve_pending_insumo_row(
        pending: PointPendingMatch,
        insumo_target: Insumo,
        create_aliases_enabled: bool,
    ) -> tuple[bool, bool, int]:
        point_code = (pending.point_codigo or "").strip()
        if point_code and insumo_target.codigo_point and insumo_target.codigo_point != point_code:
            return False, True, 0

        changed = []
        if point_code and insumo_target.codigo_point != point_code:
            insumo_target.codigo_point = point_code[:80]
            changed.append("codigo_point")
        if insumo_target.nombre_point != pending.point_nombre:
            insumo_target.nombre_point = (pending.point_nombre or "")[:250]
            changed.append("nombre_point")
        if changed:
            insumo_target.save(update_fields=changed)

        alias_created = 0
        if create_aliases_enabled:
            alias_norm = normalizar_nombre(pending.point_nombre or "")
            if alias_norm and alias_norm != insumo_target.nombre_normalizado:
                alias, was_created = InsumoAlias.objects.get_or_create(
                    nombre_normalizado=alias_norm,
                    defaults={"nombre": (pending.point_nombre or "")[:250], "insumo": insumo_target},
                )
                if not was_created and alias.insumo_id != insumo_target.id:
                    alias.insumo = insumo_target
                    alias.save(update_fields=["insumo"])
                if was_created:
                    alias_created = 1

        pending.delete()
        return True, False, alias_created

    def post(self, request):
        if not has_any_role(request.user, ROLE_ADMIN, ROLE_COMPRAS):
            return Response(
                {"detail": "No tienes permisos para resolver pendientes Point."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = InventarioPointPendingResolveSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        action = data["action"]
        tipo = data["tipo"]
        pending_ids = list(dict.fromkeys(data.get("pending_ids") or []))
        q = (data.get("q") or "").strip()
        score_min = float(data.get("score_min", 90.0) or 0)
        create_aliases = bool(data.get("create_aliases", True))

        selected_qs = PointPendingMatch.objects.filter(tipo=tipo)
        if pending_ids:
            selected_qs = selected_qs.filter(id__in=pending_ids)
        elif action == InventarioPointPendingResolveSerializer.ACTION_AUTO_RESOLVE_INSUMOS:
            if q:
                selected_qs = selected_qs.filter(
                    Q(point_nombre__icontains=q)
                    | Q(point_codigo__icontains=q)
                    | Q(fuzzy_sugerencia__icontains=q)
                )
        selected_qs = selected_qs.order_by("-fuzzy_score", "point_nombre")

        selected = list(selected_qs)
        if not selected:
            return Response(
                {"detail": "No se encontraron pendientes para procesar con los filtros enviados."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        resolved = 0
        conflicts = 0
        aliases_created = 0
        skipped_low_score = 0
        skipped_no_suggestion = 0
        skipped_no_target = 0
        proveedores_created = 0

        if action == InventarioPointPendingResolveSerializer.ACTION_RESOLVE_INSUMOS:
            insumo_target = canonical_insumo_by_id(data.get("insumo_id"))
            if insumo_target is None:
                return Response(
                    {"detail": "insumo_id no encontrado o inactivo."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            for pending in selected:
                row_resolved, row_conflict, row_alias_created = self._resolve_pending_insumo_row(
                    pending=pending,
                    insumo_target=insumo_target,
                    create_aliases_enabled=create_aliases,
                )
                if row_conflict:
                    conflicts += 1
                    continue
                if row_resolved:
                    resolved += 1
                    aliases_created += row_alias_created

            return Response(
                {
                    "action": action,
                    "tipo": tipo,
                    "selected": len(selected),
                    "resolved": resolved,
                    "conflicts": conflicts,
                    "aliases_created": aliases_created,
                    "target": {"insumo_id": insumo_target.id, "insumo": insumo_target.nombre},
                },
                status=status.HTTP_200_OK,
            )

        if action == InventarioPointPendingResolveSerializer.ACTION_AUTO_RESOLVE_INSUMOS:
            if tipo != PointPendingMatch.TIPO_INSUMO:
                return Response(
                    {"detail": "La auto-resolución por sugerencia aplica solo para tipo=INSUMO."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            for pending in selected:
                if float(pending.fuzzy_score or 0.0) < score_min:
                    skipped_low_score += 1
                    continue

                sugerencia_norm = normalizar_nombre(pending.fuzzy_sugerencia or "")
                if not sugerencia_norm:
                    skipped_no_suggestion += 1
                    continue

                insumo_target = (
                    Insumo.objects.filter(activo=True, nombre_normalizado=sugerencia_norm)
                    .only("id", "codigo_point", "nombre_point", "nombre_normalizado")
                    .first()
                )
                insumo_target = canonical_insumo(insumo_target)
                if not insumo_target:
                    skipped_no_target += 1
                    continue

                row_resolved, row_conflict, row_alias_created = self._resolve_pending_insumo_row(
                    pending=pending,
                    insumo_target=insumo_target,
                    create_aliases_enabled=create_aliases,
                )
                if row_conflict:
                    conflicts += 1
                    continue
                if row_resolved:
                    resolved += 1
                    aliases_created += row_alias_created

            return Response(
                {
                    "action": action,
                    "tipo": tipo,
                    "selected": len(selected),
                    "resolved": resolved,
                    "conflicts": conflicts,
                    "aliases_created": aliases_created,
                    "skipped": {
                        "low_score": skipped_low_score,
                        "no_suggestion": skipped_no_suggestion,
                        "no_target": skipped_no_target,
                    },
                    "score_min": round(score_min, 2),
                },
                status=status.HTTP_200_OK,
            )

        if action == InventarioPointPendingResolveSerializer.ACTION_RESOLVE_PRODUCTOS:
            receta_target = Receta.objects.filter(pk=data.get("receta_id")).first()
            if receta_target is None:
                return Response(
                    {"detail": "receta_id no encontrada."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            for pending in selected:
                point_code = (pending.point_codigo or "").strip()
                if point_code:
                    point_norm = normalizar_codigo_point(point_code)
                    primary_norm = normalizar_codigo_point(receta_target.codigo_point)
                    if not receta_target.codigo_point:
                        receta_target.codigo_point = point_code[:80]
                        receta_target.save(update_fields=["codigo_point"])
                    elif primary_norm != point_norm:
                        if not point_norm:
                            conflicts += 1
                            continue
                        if not create_aliases:
                            conflicts += 1
                            continue

                        alias, was_created = RecetaCodigoPointAlias.objects.get_or_create(
                            codigo_point_normalizado=point_norm,
                            defaults={
                                "receta": receta_target,
                                "codigo_point": point_code[:80],
                                "nombre_point": (pending.point_nombre or "")[:250],
                                "activo": True,
                            },
                        )
                        if not was_created and alias.receta_id != receta_target.id:
                            conflicts += 1
                            continue

                        changed = []
                        if alias.codigo_point != point_code[:80]:
                            alias.codigo_point = point_code[:80]
                            changed.append("codigo_point")
                        if (pending.point_nombre or "").strip() and alias.nombre_point != (pending.point_nombre or "")[:250]:
                            alias.nombre_point = (pending.point_nombre or "")[:250]
                            changed.append("nombre_point")
                        if not alias.activo:
                            alias.activo = True
                            changed.append("activo")
                        if changed:
                            alias.save(update_fields=changed)
                        if was_created:
                            aliases_created += 1

                pending.delete()
                resolved += 1

            return Response(
                {
                    "action": action,
                    "tipo": tipo,
                    "selected": len(selected),
                    "resolved": resolved,
                    "conflicts": conflicts,
                    "aliases_created": aliases_created,
                    "target": {"receta_id": receta_target.id, "receta": receta_target.nombre},
                },
                status=status.HTTP_200_OK,
            )

        if action == InventarioPointPendingResolveSerializer.ACTION_RESOLVE_PROVEEDORES:
            proveedor_target = None
            proveedor_id = data.get("proveedor_id")
            if proveedor_id:
                proveedor_target = Proveedor.objects.filter(pk=proveedor_id).first()
                if proveedor_target is None:
                    return Response(
                        {"detail": "proveedor_id no encontrado."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            for pending in selected:
                if proveedor_target is None:
                    _, was_created = Proveedor.objects.get_or_create(
                        nombre=(pending.point_nombre or "")[:200],
                        defaults={"activo": True},
                    )
                    if was_created:
                        proveedores_created += 1
                pending.delete()
                resolved += 1

            return Response(
                {
                    "action": action,
                    "tipo": tipo,
                    "selected": len(selected),
                    "resolved": resolved,
                    "proveedores_created": proveedores_created,
                    "target": (
                        {"proveedor_id": proveedor_target.id, "proveedor": proveedor_target.nombre}
                        if proveedor_target
                        else None
                    ),
                },
                status=status.HTTP_200_OK,
            )

        if action == InventarioPointPendingResolveSerializer.ACTION_DISCARD:
            deleted, _ = PointPendingMatch.objects.filter(id__in=[p.id for p in selected], tipo=tipo).delete()
            return Response(
                {
                    "action": action,
                    "tipo": tipo,
                    "selected": len(selected),
                    "discarded": deleted,
                },
                status=status.HTTP_200_OK,
            )

        return Response({"detail": "Acción no soportada."}, status=status.HTTP_400_BAD_REQUEST)


class InventarioAjustesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not can_view_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para consultar ajustes de inventario."},
                status=status.HTTP_403_FORBIDDEN,
            )

        estatus = (request.GET.get("estatus") or "").strip().upper()
        valid_status = {
            AjusteInventario.STATUS_PENDIENTE,
            AjusteInventario.STATUS_APLICADO,
            AjusteInventario.STATUS_RECHAZADO,
        }
        qs = AjusteInventario.objects.select_related("insumo", "solicitado_por", "aprobado_por").order_by("-creado_en")
        if estatus in valid_status:
            qs = qs.filter(estatus=estatus)

        limit = _parse_bounded_int(
            request.GET.get("limit", 120),
            default=120,
            min_value=1,
            max_value=500,
        )
        items = [_serialize_ajuste_row(a) for a in qs[:limit]]

        totals_qs = qs if estatus in valid_status else AjusteInventario.objects.all()
        return Response(
            {
                "filters": {"estatus": estatus if estatus in valid_status else "", "limit": limit},
                "totales": {
                    "rows": len(items),
                    "pendientes": totals_qs.filter(estatus=AjusteInventario.STATUS_PENDIENTE).count(),
                    "aplicados": totals_qs.filter(estatus=AjusteInventario.STATUS_APLICADO).count(),
                    "rechazados": totals_qs.filter(estatus=AjusteInventario.STATUS_RECHAZADO).count(),
                },
                "items": items,
            },
            status=status.HTTP_200_OK,
        )

    def post(self, request):
        if not can_manage_inventario(request.user):
            return Response(
                {"detail": "No tienes permisos para registrar ajustes de inventario."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = InventarioAjusteCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        insumo = canonical_insumo_by_id(data["insumo_id"])
        if insumo is None:
            return Response(
                {"detail": "insumo_id no encontrado o inactivo."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        aplicar_inmediato = bool(data.get("aplicar_inmediato"))
        comentario = data.get("comentario_revision") or ""
        if aplicar_inmediato and not _can_approve_ajustes(request.user):
            return Response(
                {
                    "detail": (
                        "No tienes permisos para aplicar ajustes inmediatamente. "
                        "Registra el ajuste en pendiente y solicita aprobación."
                    )
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        with transaction.atomic():
            ajuste = AjusteInventario.objects.create(
                insumo=insumo,
                cantidad_sistema=data["cantidad_sistema"],
                cantidad_fisica=data["cantidad_fisica"],
                motivo=data["motivo"],
                estatus=AjusteInventario.STATUS_PENDIENTE,
                solicitado_por=request.user,
            )
            if aplicar_inmediato:
                _apply_ajuste(ajuste, request.user, comentario=comentario)

        payload = _serialize_ajuste_row(ajuste)
        payload["aplicado"] = ajuste.estatus == AjusteInventario.STATUS_APLICADO
        return Response(payload, status=status.HTTP_201_CREATED)


class InventarioAjusteDecisionView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, ajuste_id: int):
        if not _can_approve_ajustes(request.user):
            return Response(
                {"detail": "No tienes permisos para aprobar/rechazar ajustes de inventario."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = InventarioAjusteDecisionSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        action = ser.validated_data["action"]
        comentario = ser.validated_data.get("comentario_revision") or ""

        with transaction.atomic():
            ajuste = get_object_or_404(
                AjusteInventario.objects.select_for_update(of=("self",)).select_related(
                    "insumo",
                    "solicitado_por",
                    "aprobado_por",
                ),
                pk=ajuste_id,
            )
            if action == "reject":
                if ajuste.estatus == AjusteInventario.STATUS_APLICADO:
                    return Response(
                        {"detail": "No se puede rechazar un ajuste ya aplicado."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                ajuste.estatus = AjusteInventario.STATUS_RECHAZADO
                ajuste.aprobado_por = request.user
                ajuste.aprobado_en = timezone.now()
                ajuste.aplicado_en = None
                ajuste.comentario_revision = comentario
                ajuste.save(
                    update_fields=[
                        "estatus",
                        "aprobado_por",
                        "aprobado_en",
                        "aplicado_en",
                        "comentario_revision",
                    ]
                )
                log_event(
                    request.user,
                    "REJECT",
                    "inventario.AjusteInventario",
                    ajuste.id,
                    {"folio": ajuste.folio, "estatus": ajuste.estatus, "comentario_revision": comentario},
                )
            else:
                if ajuste.estatus == AjusteInventario.STATUS_APLICADO:
                    # El SELECT con JOIN puede conservar relaciones del snapshot
                    # anterior mientras espera el lock; recarga el estado confirmado.
                    ajuste = AjusteInventario.objects.select_related(
                        "insumo",
                        "solicitado_por",
                        "aprobado_por",
                    ).get(pk=ajuste.pk)
                    return Response(
                        {
                            "detail": "El ajuste ya estaba aplicado.",
                            "item": _serialize_ajuste_row(ajuste),
                        },
                        status=status.HTTP_200_OK,
                    )
                if ajuste.estatus == AjusteInventario.STATUS_RECHAZADO:
                    return Response(
                        {"detail": "El ajuste fue rechazado y no puede aplicarse."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                _apply_ajuste(ajuste, request.user, comentario=comentario)

            payload = _serialize_ajuste_row(ajuste)
            payload["action"] = action
            return Response(payload, status=status.HTTP_200_OK)
