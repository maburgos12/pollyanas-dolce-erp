import csv
import json
from collections import defaultdict
from contextlib import nullcontext
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from django.db import transaction, OperationalError, ProgrammingError
from django.db.models import Count, Max, Q, Sum
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from openpyxl import Workbook
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.authtoken.models import Token
from rest_framework.authtoken.views import ObtainAuthToken

from activos.models import Activo, BitacoraMantenimiento, OrdenMantenimiento, PlanMantenimiento
from control.models import MermaPOS, VentaPOS
from control.services import build_discrepancias_report, resolve_period_range
from compras.models import OrdenCompra, RecepcionCompra, SolicitudCompra
from core.access import (
    ROLE_ADMIN,
    ROLE_COMPRAS,
    ROLE_DG,
    can_manage_compras,
    can_manage_inventario,
    can_view_audit,
    can_view_compras,
    can_view_inventario,
    can_view_maestros,
    can_view_reportes,
    has_any_role,
)
from core.audit import log_event
from core.models import AuditLog, Sucursal
from compras.views import (
    _apply_recepcion_to_inventario,
    _active_solicitud_statuses,
    _build_budget_context,
    _build_budget_history,
    _build_category_dashboard,
    _can_transition_orden,
    _can_transition_recepcion,
    _can_transition_solicitud,
    _build_consumo_vs_plan_dashboard,
    _default_fecha_requerida,
    _parse_date_value,
    _build_provider_dashboard,
    _filtered_solicitudes,
    _resolve_proveedor_name,
    _sanitize_consumo_ref_filter,
)
from integraciones.models import PublicApiAccessLog, PublicApiClient
from integraciones.views import _deactivate_idle_api_clients, _purge_api_logs
from inventario.models import AjusteInventario, AlmacenSyncRun, ExistenciaInsumo
from inventario.views import (
    _apply_ajuste,
    _apply_cross_filters,
    _build_cross_unified_rows,
    _export_cross_pending_csv,
    _export_cross_pending_xlsx,
    _build_pending_grouped,
    _resolve_cross_source_with_alias,
)
from maestros.models import CostoInsumo, Insumo, InsumoAlias, PointPendingMatch, Proveedor
from maestros.utils.canonical_catalog import canonical_insumo, canonical_insumo_by_id, canonicalized_active_insumos, latest_costo_canonico
from recetas.models import (
    LineaReceta,
    PlanProduccion,
    PlanProduccionItem,
    PronosticoVenta,
    Receta,
    RecetaCodigoPointAlias,
    SolicitudVenta,
    VentaHistorica,
    normalizar_codigo_point,
)
from recetas.views import (
    _build_forecast_backtest_preview,
    _build_forecast_from_history,
    _filter_forecast_result_by_confianza,
    _forecast_session_payload,
    _forecast_vs_solicitud_preview,
    _normalize_periodo_mes,
    _resolve_receta_for_sales,
    _resolve_solicitud_window,
    _resolve_sucursal_for_sales,
    _ui_to_model_alcance,
)
from recetas.utils.normalizacion import normalizar_nombre
from recetas.utils.matching import match_insumo
from recetas.utils.costeo_versionado import asegurar_version_costeo, comparativo_versiones
from ..serializers import (
    ComprasSolicitudImportConfirmSerializer,
    ComprasSolicitudImportPreviewSerializer,
    ComprasCrearOrdenSerializer,
    ComprasOrdenStatusSerializer,
    ComprasRecepcionCreateSerializer,
    ComprasRecepcionStatusSerializer,
    ComprasSolicitudCreateSerializer,
    ComprasSolicitudStatusSerializer,
    ControlMermaPosBulkSerializer,
    ControlVentaPosBulkSerializer,
    ActivosOrdenCreateSerializer,
    ActivosOrdenStatusSerializer,
    ForecastBacktestRequestSerializer,
    ForecastEstadisticoGuardarSerializer,
    ForecastEstadisticoRequestSerializer,
    IntegracionesDeactivateIdleClientsSerializer,
    IntegracionesMaintenanceRunSerializer,
    IntegracionesOperationHistoryQuerySerializer,
    IntegracionesPurgeApiLogsSerializer,
    InventarioAjusteCreateSerializer,
    InventarioAjusteDecisionSerializer,
    InventarioAliasCreateSerializer,
    InventarioCrossPendientesResolveSerializer,
    InventarioAliasMassReassignSerializer,
    MasterDuplicatesSerializer,
    MasterNormalizeSerializer,
    InventarioPointPendingResolveSerializer,
    MRPRequestSerializer,
    MRPRequerimientosRequestSerializer,
    PlanProduccionCreateSerializer,
    PlanProduccionItemCreateSerializer,
    PlanProduccionItemUpdateSerializer,
    PlanProduccionUpdateSerializer,
    PlanDesdePronosticoRequestSerializer,
    PronosticoVentaBulkSerializer,
    RecetaCostoVersionSerializer,
    SolicitudVentaAplicarForecastSerializer,
    SolicitudVentaBulkSerializer,
    SolicitudVentaUpsertSerializer,
    VentaHistoricaBulkSerializer,
)


def _load_versiones_costeo(receta: Receta, limit: int):
    return list(receta.versiones_costo.order_by("-version_num")[:limit])


def _to_decimal(value, default: Decimal = Decimal("0")) -> Decimal:
    if value is None:
        return default
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return default


def _canonical_member_ids(insumo_id: int | str | None) -> tuple[Insumo | None, list[int]]:
    canonical = canonical_insumo_by_id(insumo_id)
    if canonical is None:
        return None, []
    for row in canonicalized_active_insumos(limit=5000):
        if canonical.id in row["member_ids"]:
            return canonical, list(row["member_ids"])
    return canonical, [canonical.id]


def _latest_cost_for_canonical(insumo_id: int | str | None, proveedor: Proveedor | None = None) -> tuple[Insumo | None, CostoInsumo | None]:
    canonical, member_ids = _canonical_member_ids(insumo_id)
    if canonical is None:
        return None, None
    costo_qs = CostoInsumo.objects.filter(insumo_id__in=member_ids).order_by("-fecha", "-id")
    if proveedor is not None:
        preferred = costo_qs.filter(proveedor=proveedor).first()
        if preferred is not None:
            return canonical, preferred
    return canonical, costo_qs.first()


def _parse_bool(value: str | None, default: bool = False) -> bool:
    if value is None:
        return default
    raw = str(value).strip().lower()
    if raw in {"1", "true", "yes", "si", "on"}:
        return True
    if raw in {"0", "false", "no", "off"}:
        return False
    return default


def _parse_period(period_raw: str | None) -> tuple[int, int] | None:
    if not period_raw:
        return None
    raw = str(period_raw).strip()
    parts = raw.split("-")
    if len(parts) != 2:
        return None
    try:
        year = int(parts[0])
        month = int(parts[1])
    except ValueError:
        return None
    if year < 2000 or year > 2200 or month < 1 or month > 12:
        return None
    return year, month


def _parse_iso_date(value: str | None) -> date | None:
    if not value:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError:
        return None


def _parse_bounded_int(raw_value, *, default: int, min_value: int, max_value: int) -> int:
    try:
        value = int(raw_value)
    except (TypeError, ValueError):
        value = default
    return max(min_value, min(value, max_value))


def _to_float(value, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError, InvalidOperation):
        return default


def _pct_change(current: int, previous: int) -> float:
    current_i = int(current or 0)
    previous_i = int(previous or 0)
    if previous_i <= 0:
        return 100.0 if current_i > 0 else 0.0
    return round(((current_i - previous_i) * 100.0 / previous_i), 2)


def _build_public_api_daily_trend(days: int = 7) -> list[dict[str, Any]]:
    days = max(1, min(int(days or 7), 31))
    today = timezone.localdate()
    start_date = today - timedelta(days=days - 1)
    raw_rows = list(
        PublicApiAccessLog.objects.filter(created_at__date__gte=start_date)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(
            total=Count("id"),
            errors=Count("id", filter=Q(status_code__gte=400)),
        )
        .order_by("day")
    )
    by_day = {row["day"]: row for row in raw_rows}
    trend = []
    for index in range(days):
        day = start_date + timedelta(days=index)
        row = by_day.get(day, {})
        total = int(row.get("total") or 0)
        errors = int(row.get("errors") or 0)
        trend.append(
            {
                "day": day,
                "requests": total,
                "errors": errors,
                "error_rate_pct": round((errors * 100.0 / total), 2) if total else 0.0,
            }
        )
    return trend


def _preview_deactivate_idle_api_clients(idle_days: int, limit: int) -> dict[str, Any]:
    idle_days = max(1, min(int(idle_days or 30), 365))
    limit = max(1, min(int(limit or 100), 500))
    cutoff = timezone.now() - timedelta(days=idle_days)
    recent_client_ids = set(
        PublicApiAccessLog.objects.filter(created_at__gte=cutoff)
        .values_list("client_id", flat=True)
        .distinct()
    )
    candidates = list(
        PublicApiClient.objects.filter(activo=True)
        .exclude(id__in=recent_client_ids)
        .order_by("id")
        .values("id", "nombre")[:limit]
    )
    return {
        "idle_days": idle_days,
        "limit": limit,
        "candidates": len(candidates),
        "deactivated": 0,
        "cutoff": cutoff.isoformat(),
        "candidate_clients": candidates,
        "dry_run": True,
    }


def _preview_purge_api_logs(retain_days: int, max_delete: int) -> dict[str, Any]:
    retain_days = max(1, min(int(retain_days or 90), 3650))
    max_delete = max(1, min(int(max_delete or 5000), 50000))
    cutoff = timezone.now() - timedelta(days=retain_days)
    total_candidates = PublicApiAccessLog.objects.filter(created_at__lt=cutoff).count()
    return {
        "retain_days": retain_days,
        "max_delete": max_delete,
        "cutoff": cutoff.isoformat(),
        "candidates": int(total_candidates),
        "deleted": 0,
        "remaining_candidates": int(total_candidates),
        "would_delete": min(int(total_candidates), max_delete),
        "dry_run": True,
    }



def _serialize_forecast_compare(compare: dict | None, *, top: int = 120) -> dict:
    if not compare:
        return {
            "target_start": "",
            "target_end": "",
            "sucursal_id": None,
            "sucursal_nombre": "",
            "escenario": "base",
            "rows": [],
            "totals": {
                "forecast_total": 0.0,
                "solicitud_total": 0.0,
                "delta_total": 0.0,
                "ok_count": 0,
                "sobre_count": 0,
                "bajo_count": 0,
                "sin_base_count": 0,
                "en_rango_count": 0,
                "sobre_rango_count": 0,
                "bajo_rango_count": 0,
            },
        }

    rows = []
    for row in (compare.get("rows") or [])[:top]:
        variacion = row.get("variacion_pct")
        rows.append(
            {
                "receta_id": int(row.get("receta_id") or 0),
                "receta": row.get("receta") or "",
                "forecast_qty": _to_float(row.get("forecast_qty")),
                "forecast_base": _to_float(row.get("forecast_base")),
                "forecast_low": _to_float(row.get("forecast_low")),
                "forecast_high": _to_float(row.get("forecast_high")),
                "solicitud_qty": _to_float(row.get("solicitud_qty")),
                "delta_qty": _to_float(row.get("delta_qty")),
                "variacion_pct": _to_float(variacion) if variacion is not None else None,
                "status": row.get("status") or "",
                "status_rango": row.get("status_rango") or "",
            }
        )
    totals = compare.get("totals") or {}
    return {
        "target_start": str(compare.get("target_start") or ""),
        "target_end": str(compare.get("target_end") or ""),
        "sucursal_id": compare.get("sucursal_id"),
        "sucursal_nombre": compare.get("sucursal_nombre") or "",
        "escenario": compare.get("escenario") or "base",
        "rows": rows,
        "totals": {
            "forecast_total": _to_float(totals.get("forecast_total")),
            "solicitud_total": _to_float(totals.get("solicitud_total")),
            "delta_total": _to_float(totals.get("delta_total")),
            "ok_count": int(totals.get("ok_count") or 0),
            "sobre_count": int(totals.get("sobre_count") or 0),
            "bajo_count": int(totals.get("bajo_count") or 0),
            "sin_base_count": int(totals.get("sin_base_count") or 0),
            "en_rango_count": int(totals.get("en_rango_count") or 0),
            "sobre_rango_count": int(totals.get("sobre_rango_count") or 0),
            "bajo_rango_count": int(totals.get("bajo_rango_count") or 0),
        },
    }


def _ventas_pipeline_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    scope = payload.get("scope") or {}
    totals = payload.get("totales") or {}
    by_alcance = payload.get("solicitud_by_alcance") or {}
    by_sucursal = payload.get("by_sucursal") or []
    by_sucursal_status = totals.get("by_sucursal_status") or {}
    rows_status = totals.get("rows_status") or {}
    rows = payload.get("rows") or []
    periodo = str(scope.get("periodo") or _normalize_periodo_mes(None)).replace("-", "")
    filename = f"ventas_pipeline_{periodo}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Periodo", scope.get("periodo") or ""])
        ws_resumen.append(["Sucursal", scope.get("sucursal") or "Todas"])
        ws_resumen.append(["Incluir preparaciones", "SI" if scope.get("incluir_preparaciones") else "NO"])
        ws_resumen.append(["Filtro q", scope.get("q") or ""])
        ws_resumen.append(["Top", int(scope.get("top") or 0)])
        ws_resumen.append(["Offset", int(scope.get("offset") or 0)])
        ws_resumen.append(["Top sucursales", int(scope.get("top_sucursales") or 0)])
        ws_resumen.append(["Offset sucursales", int(scope.get("offset_sucursales") or 0)])
        ws_resumen.append(["Sort rows", str(scope.get("sort_by") or "delta_abs").upper()])
        ws_resumen.append(["Sort rows dir", str(scope.get("sort_dir") or "desc").upper()])
        ws_resumen.append(["Sort sucursales", str(scope.get("sort_sucursales_by") or "delta_abs").upper()])
        ws_resumen.append(["Sort sucursales dir", str(scope.get("sort_sucursales_dir") or "desc").upper()])
        ws_resumen.append(["Rows filtered", int(totals.get("rows_filtered") or 0)])
        ws_resumen.append(["Rows returned", int(totals.get("rows_returned") or 0)])
        ws_resumen.append(["Sucursales filtered", int(totals.get("by_sucursal_filtered") or 0)])
        ws_resumen.append(["Sucursales returned", int(totals.get("by_sucursal_returned") or 0)])
        ws_resumen.append(["Historial total", float(totals.get("historial_qty") or 0)])
        ws_resumen.append(["Pronostico total", float(totals.get("pronostico_qty") or 0)])
        ws_resumen.append(["Solicitud total", float(totals.get("solicitud_qty") or 0)])
        ws_resumen.append(["Delta solicitud vs pronostico", float(totals.get("delta_solicitud_vs_pronostico") or 0)])
        ws_resumen.append(["Delta historial vs solicitud", float(totals.get("delta_historial_vs_solicitud") or 0)])
        ws_resumen.append(["Cobertura solicitud %", float(totals.get("cobertura_solicitud_pct")) if totals.get("cobertura_solicitud_pct") is not None else None])
        ws_resumen.append(["Cumplimiento historial %", float(totals.get("cumplimiento_historial_pct")) if totals.get("cumplimiento_historial_pct") is not None else None])
        ws_resumen.append(["Solicitudes MES", float(by_alcance.get("MES") or 0)])
        ws_resumen.append(["Solicitudes SEMANA", float(by_alcance.get("SEMANA") or 0)])
        ws_resumen.append(["Solicitudes FIN_SEMANA", float(by_alcance.get("FIN_SEMANA") or 0)])
        ws_resumen.append(["Rows status SOBRE", int(rows_status.get("SOBRE") or 0)])
        ws_resumen.append(["Rows status BAJO", int(rows_status.get("BAJO") or 0)])
        ws_resumen.append(["Rows status OK", int(rows_status.get("OK") or 0)])
        ws_resumen.append(["Rows status SIN_SOLICITUD", int(rows_status.get("SIN_SOLICITUD") or 0)])
        ws_resumen.append(["Rows status SIN_MOV", int(rows_status.get("SIN_MOV") or 0)])
        ws_resumen.append(["Sucursales status SOBRE", int(by_sucursal_status.get("SOBRE") or 0)])
        ws_resumen.append(["Sucursales status BAJO", int(by_sucursal_status.get("BAJO") or 0)])
        ws_resumen.append(["Sucursales status OK", int(by_sucursal_status.get("OK") or 0)])
        ws_resumen.append(["Sucursales status SIN_SOLICITUD", int(by_sucursal_status.get("SIN_SOLICITUD") or 0)])
        ws_resumen.append(["Sucursales status SIN_MOV", int(by_sucursal_status.get("SIN_MOV") or 0)])

        ws_sucursal = wb.create_sheet("BySucursal")
        ws_sucursal.append(
            [
                "Sucursal ID",
                "Sucursal codigo",
                "Sucursal",
                "Historial",
                "Solicitud",
                "Delta Historial vs Solicitud",
                "Cumplimiento %",
                "Status",
            ]
        )
        for row in by_sucursal:
            ws_sucursal.append(
                [
                    int(row.get("sucursal_id") or 0),
                    row.get("sucursal_codigo") or "",
                    row.get("sucursal") or "",
                    float(row.get("historial_qty") or 0),
                    float(row.get("solicitud_qty") or 0),
                    float(row.get("delta_historial_vs_solicitud") or 0),
                    float(row.get("cumplimiento_pct")) if row.get("cumplimiento_pct") is not None else None,
                    row.get("status") or "",
                ]
            )

        ws_detalle = wb.create_sheet("Detalle")
        ws_detalle.append(
            [
                "Receta ID",
                "Receta",
                "Historial",
                "Pronostico",
                "Solicitud",
                "Delta Solicitud vs Pronostico",
                "Delta Historial vs Solicitud",
                "Cobertura %",
                "Cumplimiento %",
                "Status",
            ]
        )
        for row in rows:
            ws_detalle.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    float(row.get("historial_qty") or 0),
                    float(row.get("pronostico_qty") or 0),
                    float(row.get("solicitud_qty") or 0),
                    float(row.get("delta_solicitud_vs_pronostico") or 0),
                    float(row.get("delta_historial_vs_solicitud") or 0),
                    float(row.get("cobertura_pct")) if row.get("cobertura_pct") is not None else None,
                    float(row.get("cumplimiento_pct")) if row.get("cumplimiento_pct") is not None else None,
                    row.get("status") or "",
                ]
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Periodo", scope.get("periodo") or ""])
    writer.writerow(["Sucursal", scope.get("sucursal") or "Todas"])
    writer.writerow(["Incluir preparaciones", "SI" if scope.get("incluir_preparaciones") else "NO"])
    writer.writerow(["Filtro q", scope.get("q") or ""])
    writer.writerow(["Top", int(scope.get("top") or 0)])
    writer.writerow(["Offset", int(scope.get("offset") or 0)])
    writer.writerow(["Top sucursales", int(scope.get("top_sucursales") or 0)])
    writer.writerow(["Offset sucursales", int(scope.get("offset_sucursales") or 0)])
    writer.writerow(["Sort rows", str(scope.get("sort_by") or "delta_abs").upper()])
    writer.writerow(["Sort rows dir", str(scope.get("sort_dir") or "desc").upper()])
    writer.writerow(["Sort sucursales", str(scope.get("sort_sucursales_by") or "delta_abs").upper()])
    writer.writerow(["Sort sucursales dir", str(scope.get("sort_sucursales_dir") or "desc").upper()])
    writer.writerow(["Rows filtered", int(totals.get("rows_filtered") or 0)])
    writer.writerow(["Rows returned", int(totals.get("rows_returned") or 0)])
    writer.writerow(["Sucursales filtered", int(totals.get("by_sucursal_filtered") or 0)])
    writer.writerow(["Sucursales returned", int(totals.get("by_sucursal_returned") or 0)])
    writer.writerow(["Historial total", f"{Decimal(str(totals.get('historial_qty') or 0)):.3f}"])
    writer.writerow(["Pronostico total", f"{Decimal(str(totals.get('pronostico_qty') or 0)):.3f}"])
    writer.writerow(["Solicitud total", f"{Decimal(str(totals.get('solicitud_qty') or 0)):.3f}"])
    writer.writerow(["Delta solicitud vs pronostico", f"{Decimal(str(totals.get('delta_solicitud_vs_pronostico') or 0)):.3f}"])
    writer.writerow(["Delta historial vs solicitud", f"{Decimal(str(totals.get('delta_historial_vs_solicitud') or 0)):.3f}"])
    writer.writerow(
        [
            "Cobertura solicitud %",
            f"{Decimal(str(totals.get('cobertura_solicitud_pct'))):.1f}" if totals.get("cobertura_solicitud_pct") is not None else "",
        ]
    )
    writer.writerow(
        [
            "Cumplimiento historial %",
            f"{Decimal(str(totals.get('cumplimiento_historial_pct'))):.1f}" if totals.get("cumplimiento_historial_pct") is not None else "",
        ]
    )
    writer.writerow(["Solicitudes MES", f"{Decimal(str(by_alcance.get('MES') or 0)):.3f}"])
    writer.writerow(["Solicitudes SEMANA", f"{Decimal(str(by_alcance.get('SEMANA') or 0)):.3f}"])
    writer.writerow(["Solicitudes FIN_SEMANA", f"{Decimal(str(by_alcance.get('FIN_SEMANA') or 0)):.3f}"])
    writer.writerow(["Rows status SOBRE", int(rows_status.get("SOBRE") or 0)])
    writer.writerow(["Rows status BAJO", int(rows_status.get("BAJO") or 0)])
    writer.writerow(["Rows status OK", int(rows_status.get("OK") or 0)])
    writer.writerow(["Rows status SIN_SOLICITUD", int(rows_status.get("SIN_SOLICITUD") or 0)])
    writer.writerow(["Rows status SIN_MOV", int(rows_status.get("SIN_MOV") or 0)])
    writer.writerow(["Sucursales status SOBRE", int(by_sucursal_status.get("SOBRE") or 0)])
    writer.writerow(["Sucursales status BAJO", int(by_sucursal_status.get("BAJO") or 0)])
    writer.writerow(["Sucursales status OK", int(by_sucursal_status.get("OK") or 0)])
    writer.writerow(["Sucursales status SIN_SOLICITUD", int(by_sucursal_status.get("SIN_SOLICITUD") or 0)])
    writer.writerow(["Sucursales status SIN_MOV", int(by_sucursal_status.get("SIN_MOV") or 0)])
    writer.writerow([])
    writer.writerow(["BY_SUCURSAL"])
    writer.writerow(
        [
            "sucursal_id",
            "sucursal_codigo",
            "sucursal",
            "historial_qty",
            "solicitud_qty",
            "delta_historial_vs_solicitud",
            "cumplimiento_pct",
            "status",
        ]
    )
    for row in by_sucursal:
        writer.writerow(
            [
                int(row.get("sucursal_id") or 0),
                row.get("sucursal_codigo") or "",
                row.get("sucursal") or "",
                f"{Decimal(str(row.get('historial_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('solicitud_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('delta_historial_vs_solicitud') or 0)):.3f}",
                f"{Decimal(str(row.get('cumplimiento_pct'))):.1f}" if row.get("cumplimiento_pct") is not None else "",
                row.get("status") or "",
            ]
        )
    writer.writerow([])
    writer.writerow(
        [
            "receta_id",
            "receta",
            "historial_qty",
            "pronostico_qty",
            "solicitud_qty",
            "delta_solicitud_vs_pronostico",
            "delta_historial_vs_solicitud",
            "cobertura_pct",
            "cumplimiento_pct",
            "status",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                f"{Decimal(str(row.get('historial_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('pronostico_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('solicitud_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('delta_solicitud_vs_pronostico') or 0)):.3f}",
                f"{Decimal(str(row.get('delta_historial_vs_solicitud') or 0)):.3f}",
                f"{Decimal(str(row.get('cobertura_pct'))):.1f}" if row.get("cobertura_pct") is not None else "",
                f"{Decimal(str(row.get('cumplimiento_pct'))):.1f}" if row.get("cumplimiento_pct") is not None else "",
                row.get("status") or "",
            ]
        )
    return response


def _forecast_estadistico_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    scope = payload.get("scope") or {}
    totals = payload.get("totals") or {}
    rows = payload.get("rows") or []
    compare = payload.get("compare_solicitud") or {}
    include_compare = bool(compare.get("rows"))
    start_txt = str(scope.get("target_start") or "").replace("-", "")
    end_txt = str(scope.get("target_end") or "").replace("-", "")
    if not start_txt or not end_txt:
        today_txt = timezone.localdate().strftime("%Y%m%d")
        start_txt = start_txt or today_txt
        end_txt = end_txt or today_txt
    filename = f"api_forecast_estadistico_{start_txt}_{end_txt}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Alcance", str(scope.get("alcance") or "").upper()])
        ws_resumen.append(["Periodo", scope.get("periodo") or ""])
        ws_resumen.append(["Rango inicio", scope.get("target_start") or ""])
        ws_resumen.append(["Rango fin", scope.get("target_end") or ""])
        ws_resumen.append(["Sucursal", scope.get("sucursal_nombre") or "Todas"])
        ws_resumen.append(["Escenario compare", str(scope.get("escenario_compare") or "base").upper()])
        ws_resumen.append(["Confianza minima %", float(scope.get("min_confianza_pct") or 0)])
        ws_resumen.append(["Forecast total", float(totals.get("forecast_total") or 0)])
        ws_resumen.append(["Banda baja total", float(totals.get("forecast_low_total") or 0)])
        ws_resumen.append(["Banda alta total", float(totals.get("forecast_high_total") or 0)])
        ws_resumen.append(["Pronostico total", float(totals.get("pronostico_total") or 0)])
        ws_resumen.append(["Delta total", float(totals.get("delta_total") or 0)])

        ws_forecast = wb.create_sheet("Forecast")
        ws_forecast.append(
            [
                "Receta ID",
                "Receta",
                "Forecast",
                "Banda baja",
                "Banda alta",
                "Pronostico actual",
                "Delta",
                "Recomendacion",
                "Confianza %",
                "Desviacion",
                "Muestras",
                "Observaciones",
            ]
        )
        for row in rows:
            ws_forecast.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    float(row.get("forecast_qty") or 0),
                    float(row.get("forecast_low") or 0),
                    float(row.get("forecast_high") or 0),
                    float(row.get("pronostico_actual") or 0),
                    float(row.get("delta") or 0),
                    row.get("recomendacion") or "",
                    float(row.get("confianza") or 0),
                    float(row.get("desviacion") or 0),
                    int(row.get("muestras") or 0),
                    row.get("observaciones") or "",
                ]
            )

        if include_compare:
            ws_compare = wb.create_sheet("CompareSolicitud")
            ws_compare.append(
                [
                    "Receta ID",
                    "Receta",
                    "Forecast",
                    "Forecast base",
                    "Forecast baja",
                    "Forecast alta",
                    "Solicitud",
                    "Delta",
                    "Variacion %",
                    "Status",
                    "Status rango",
                ]
            )
            for row in compare.get("rows") or []:
                ws_compare.append(
                    [
                        int(row.get("receta_id") or 0),
                        row.get("receta") or "",
                        float(row.get("forecast_qty") or 0),
                        float(row.get("forecast_base") or 0),
                        float(row.get("forecast_low") or 0),
                        float(row.get("forecast_high") or 0),
                        float(row.get("solicitud_qty") or 0),
                        float(row.get("delta_qty") or 0),
                        float(row.get("variacion_pct")) if row.get("variacion_pct") is not None else None,
                        row.get("status") or "",
                        row.get("status_rango") or "",
                    ]
                )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Alcance", str(scope.get("alcance") or "").upper()])
    writer.writerow(["Periodo", scope.get("periodo") or ""])
    writer.writerow(["Rango inicio", scope.get("target_start") or ""])
    writer.writerow(["Rango fin", scope.get("target_end") or ""])
    writer.writerow(["Sucursal", scope.get("sucursal_nombre") or "Todas"])
    writer.writerow(["Escenario compare", str(scope.get("escenario_compare") or "base").upper()])
    writer.writerow(["Confianza minima %", f"{Decimal(str(scope.get('min_confianza_pct') or 0)):.1f}"])
    writer.writerow(["Forecast total", f"{Decimal(str(totals.get('forecast_total') or 0)):.3f}"])
    writer.writerow(["Banda baja total", f"{Decimal(str(totals.get('forecast_low_total') or 0)):.3f}"])
    writer.writerow(["Banda alta total", f"{Decimal(str(totals.get('forecast_high_total') or 0)):.3f}"])
    writer.writerow(["Pronostico total", f"{Decimal(str(totals.get('pronostico_total') or 0)):.3f}"])
    writer.writerow(["Delta total", f"{Decimal(str(totals.get('delta_total') or 0)):.3f}"])
    writer.writerow([])
    writer.writerow(
        [
            "receta_id",
            "receta",
            "forecast",
            "banda_baja",
            "banda_alta",
            "pronostico_actual",
            "delta",
            "recomendacion",
            "confianza_pct",
            "desviacion",
            "muestras",
            "observaciones",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                f"{Decimal(str(row.get('forecast_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_low') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_high') or 0)):.3f}",
                f"{Decimal(str(row.get('pronostico_actual') or 0)):.3f}",
                f"{Decimal(str(row.get('delta') or 0)):.3f}",
                row.get("recomendacion") or "",
                f"{Decimal(str(row.get('confianza') or 0)):.1f}",
                f"{Decimal(str(row.get('desviacion') or 0)):.3f}",
                int(row.get("muestras") or 0),
                row.get("observaciones") or "",
            ]
        )
    if include_compare:
        writer.writerow([])
        writer.writerow(["COMPARE_SOLICITUD"])
        writer.writerow(
            [
                "receta_id",
                "receta",
                "forecast",
                "forecast_base",
                "forecast_baja",
                "forecast_alta",
                "solicitud",
                "delta",
                "variacion_pct",
                "status",
                "status_rango",
            ]
        )
        for row in compare.get("rows") or []:
            writer.writerow(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    f"{Decimal(str(row.get('forecast_qty') or 0)):.3f}",
                    f"{Decimal(str(row.get('forecast_base') or 0)):.3f}",
                    f"{Decimal(str(row.get('forecast_low') or 0)):.3f}",
                    f"{Decimal(str(row.get('forecast_high') or 0)):.3f}",
                    f"{Decimal(str(row.get('solicitud_qty') or 0)):.3f}",
                    f"{Decimal(str(row.get('delta_qty') or 0)):.3f}",
                    f"{Decimal(str(row.get('variacion_pct'))):.1f}" if row.get("variacion_pct") is not None else "",
                    row.get("status") or "",
                    row.get("status_rango") or "",
                ]
            )
    return response


def _forecast_estadistico_guardar_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    scope = payload.get("scope") or {}
    totals = payload.get("totals") or {}
    persisted = payload.get("persisted") or {}
    rows = payload.get("rows") or []
    applied_rows = payload.get("applied_rows") or []
    start_txt = str(scope.get("target_start") or "").replace("-", "")
    end_txt = str(scope.get("target_end") or "").replace("-", "")
    if not start_txt or not end_txt:
        today_txt = timezone.localdate().strftime("%Y%m%d")
        start_txt = start_txt or today_txt
        end_txt = end_txt or today_txt
    filename = f"api_forecast_estadistico_guardar_{start_txt}_{end_txt}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Alcance", str(scope.get("alcance") or "").upper()])
        ws_resumen.append(["Periodo", scope.get("periodo") or ""])
        ws_resumen.append(["Rango inicio", scope.get("target_start") or ""])
        ws_resumen.append(["Rango fin", scope.get("target_end") or ""])
        ws_resumen.append(["Sucursal", scope.get("sucursal_nombre") or "Todas"])
        ws_resumen.append(["Escenario", str(scope.get("escenario") or "base").upper()])
        ws_resumen.append(["Confianza minima %", float(scope.get("min_confianza_pct") or 0)])
        ws_resumen.append(["Filtradas por confianza", int(scope.get("filtered_conf") or 0)])
        ws_resumen.append(["Forecast total", float(totals.get("forecast_total") or 0)])
        ws_resumen.append(["Banda baja total", float(totals.get("forecast_low_total") or 0)])
        ws_resumen.append(["Banda alta total", float(totals.get("forecast_high_total") or 0)])
        ws_resumen.append(["Pronostico total", float(totals.get("pronostico_total") or 0)])
        ws_resumen.append(["Delta total", float(totals.get("delta_total") or 0)])
        ws_resumen.append(["Created", int(persisted.get("created") or 0)])
        ws_resumen.append(["Updated", int(persisted.get("updated") or 0)])
        ws_resumen.append(["Skipped existing", int(persisted.get("skipped_existing") or 0)])
        ws_resumen.append(["Skipped invalid", int(persisted.get("skipped_invalid") or 0)])
        ws_resumen.append(["Applied", int(persisted.get("applied") or 0)])

        ws_forecast = wb.create_sheet("Forecast")
        ws_forecast.append(
            [
                "Receta ID",
                "Receta",
                "Forecast",
                "Banda baja",
                "Banda alta",
                "Pronostico actual",
                "Delta",
                "Recomendacion",
                "Confianza %",
                "Desviacion",
                "Muestras",
                "Observaciones",
            ]
        )
        for row in rows:
            ws_forecast.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    float(row.get("forecast_qty") or 0),
                    float(row.get("forecast_low") or 0),
                    float(row.get("forecast_high") or 0),
                    float(row.get("pronostico_actual") or 0),
                    float(row.get("delta") or 0),
                    row.get("recomendacion") or "",
                    float(row.get("confianza") or 0),
                    float(row.get("desviacion") or 0),
                    int(row.get("muestras") or 0),
                    row.get("observaciones") or "",
                ]
            )

        ws_applied = wb.create_sheet("Applied")
        ws_applied.append(["Receta ID", "Receta", "Escenario", "Cantidad anterior", "Cantidad nueva", "Accion"])
        for row in applied_rows:
            ws_applied.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    row.get("escenario") or "",
                    float(row.get("cantidad_anterior") or 0),
                    float(row.get("cantidad_nueva") or 0),
                    row.get("accion") or "",
                ]
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Alcance", str(scope.get("alcance") or "").upper()])
    writer.writerow(["Periodo", scope.get("periodo") or ""])
    writer.writerow(["Rango inicio", scope.get("target_start") or ""])
    writer.writerow(["Rango fin", scope.get("target_end") or ""])
    writer.writerow(["Sucursal", scope.get("sucursal_nombre") or "Todas"])
    writer.writerow(["Escenario", str(scope.get("escenario") or "base").upper()])
    writer.writerow(["Confianza minima %", f"{Decimal(str(scope.get('min_confianza_pct') or 0)):.1f}"])
    writer.writerow(["Filtradas por confianza", int(scope.get("filtered_conf") or 0)])
    writer.writerow(["Forecast total", f"{Decimal(str(totals.get('forecast_total') or 0)):.3f}"])
    writer.writerow(["Banda baja total", f"{Decimal(str(totals.get('forecast_low_total') or 0)):.3f}"])
    writer.writerow(["Banda alta total", f"{Decimal(str(totals.get('forecast_high_total') or 0)):.3f}"])
    writer.writerow(["Pronostico total", f"{Decimal(str(totals.get('pronostico_total') or 0)):.3f}"])
    writer.writerow(["Delta total", f"{Decimal(str(totals.get('delta_total') or 0)):.3f}"])
    writer.writerow(["Created", int(persisted.get("created") or 0)])
    writer.writerow(["Updated", int(persisted.get("updated") or 0)])
    writer.writerow(["Skipped existing", int(persisted.get("skipped_existing") or 0)])
    writer.writerow(["Skipped invalid", int(persisted.get("skipped_invalid") or 0)])
    writer.writerow(["Applied", int(persisted.get("applied") or 0)])
    writer.writerow([])
    writer.writerow(["FORECAST"])
    writer.writerow(
        [
            "receta_id",
            "receta",
            "forecast",
            "banda_baja",
            "banda_alta",
            "pronostico_actual",
            "delta",
            "recomendacion",
            "confianza_pct",
            "desviacion",
            "muestras",
            "observaciones",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                f"{Decimal(str(row.get('forecast_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_low') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_high') or 0)):.3f}",
                f"{Decimal(str(row.get('pronostico_actual') or 0)):.3f}",
                f"{Decimal(str(row.get('delta') or 0)):.3f}",
                row.get("recomendacion") or "",
                f"{Decimal(str(row.get('confianza') or 0)):.1f}",
                f"{Decimal(str(row.get('desviacion') or 0)):.3f}",
                int(row.get("muestras") or 0),
                row.get("observaciones") or "",
            ]
        )
    writer.writerow([])
    writer.writerow(["APPLIED"])
    writer.writerow(["receta_id", "receta", "escenario", "cantidad_anterior", "cantidad_nueva", "accion"])
    for row in applied_rows:
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                row.get("escenario") or "",
                f"{Decimal(str(row.get('cantidad_anterior') or 0)):.3f}",
                f"{Decimal(str(row.get('cantidad_nueva') or 0)):.3f}",
                row.get("accion") or "",
            ]
        )
    return response


def _forecast_backtest_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    scope = payload.get("scope") or {}
    totals = payload.get("totals") or {}
    windows = payload.get("windows") or []
    alcance = str(scope.get("alcance") or "mes").lower()
    fecha_base = str(scope.get("fecha_base") or timezone.localdate().isoformat()).replace("-", "")
    scenario = str(scope.get("escenario") or "base").lower()
    filename = f"api_forecast_backtest_{alcance}_{fecha_base}_{scenario}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Alcance", str(scope.get("alcance") or "").upper()])
        ws_resumen.append(["Fecha base", str(scope.get("fecha_base") or "")])
        ws_resumen.append(["Escenario", str(scope.get("escenario") or "base").upper()])
        ws_resumen.append(["Sucursal", scope.get("sucursal_nombre") or "Todas"])
        ws_resumen.append(["Confianza minima %", float(scope.get("min_confianza_pct") or 0)])
        ws_resumen.append(["Ventanas evaluadas", int(totals.get("windows_evaluated") or 0)])
        ws_resumen.append(["Forecast total", float(totals.get("forecast_total") or 0)])
        ws_resumen.append(["Real total", float(totals.get("actual_total") or 0)])
        ws_resumen.append(["Bias total", float(totals.get("bias_total") or 0)])
        ws_resumen.append(["MAE promedio", float(totals.get("mae_promedio") or 0)])
        ws_resumen.append(["MAPE promedio", float(totals.get("mape_promedio")) if totals.get("mape_promedio") is not None else None])

        ws_windows = wb.create_sheet("Ventanas")
        ws_windows.append(["Inicio", "Fin", "Periodo", "Recetas", "Forecast", "Real", "Bias", "MAE", "MAPE"])
        for w in windows:
            ws_windows.append(
                [
                    w.get("window_start") or "",
                    w.get("window_end") or "",
                    w.get("periodo") or "",
                    int(w.get("recetas_count") or 0),
                    float(w.get("forecast_total") or 0),
                    float(w.get("actual_total") or 0),
                    float(w.get("bias_total") or 0),
                    float(w.get("mae") or 0),
                    float(w.get("mape")) if w.get("mape") is not None else None,
                ]
            )

        ws_top = wb.create_sheet("TopErrores")
        ws_top.append(
            [
                "Inicio",
                "Fin",
                "Periodo",
                "Receta ID",
                "Receta",
                "Forecast",
                "Real",
                "Delta",
                "Abs Error",
                "Variacion %",
                "Status",
            ]
        )
        for w in windows:
            for row in w.get("top_errors") or []:
                ws_top.append(
                    [
                        w.get("window_start") or "",
                        w.get("window_end") or "",
                        w.get("periodo") or "",
                        int(row.get("receta_id") or 0),
                        row.get("receta") or "",
                        float(row.get("forecast_qty") or 0),
                        float(row.get("actual_qty") or 0),
                        float(row.get("delta_qty") or 0),
                        float(row.get("abs_error") or 0),
                        float(row.get("variacion_pct")) if row.get("variacion_pct") is not None else None,
                        row.get("status") or "",
                    ]
                )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Alcance", str(scope.get("alcance") or "").upper()])
    writer.writerow(["Fecha base", str(scope.get("fecha_base") or "")])
    writer.writerow(["Escenario", str(scope.get("escenario") or "base").upper()])
    writer.writerow(["Sucursal", scope.get("sucursal_nombre") or "Todas"])
    writer.writerow(["Confianza minima %", f"{Decimal(str(scope.get('min_confianza_pct') or 0)):.1f}"])
    writer.writerow(["Ventanas evaluadas", int(totals.get("windows_evaluated") or 0)])
    writer.writerow(["Forecast total", f"{Decimal(str(totals.get('forecast_total') or 0)):.3f}"])
    writer.writerow(["Real total", f"{Decimal(str(totals.get('actual_total') or 0)):.3f}"])
    writer.writerow(["Bias total", f"{Decimal(str(totals.get('bias_total') or 0)):.3f}"])
    writer.writerow(["MAE promedio", f"{Decimal(str(totals.get('mae_promedio') or 0)):.3f}"])
    writer.writerow(
        [
            "MAPE promedio",
            f"{Decimal(str(totals.get('mape_promedio'))):.1f}" if totals.get("mape_promedio") is not None else "",
        ]
    )
    writer.writerow([])
    writer.writerow(["VENTANAS"])
    writer.writerow(["inicio", "fin", "periodo", "recetas", "forecast", "real", "bias", "mae", "mape"])
    for w in windows:
        writer.writerow(
            [
                w.get("window_start") or "",
                w.get("window_end") or "",
                w.get("periodo") or "",
                int(w.get("recetas_count") or 0),
                f"{Decimal(str(w.get('forecast_total') or 0)):.3f}",
                f"{Decimal(str(w.get('actual_total') or 0)):.3f}",
                f"{Decimal(str(w.get('bias_total') or 0)):.3f}",
                f"{Decimal(str(w.get('mae') or 0)):.3f}",
                f"{Decimal(str(w.get('mape'))):.1f}" if w.get("mape") is not None else "",
            ]
        )
    writer.writerow([])
    writer.writerow(["TOP_ERRORES"])
    writer.writerow(
        [
            "inicio",
            "fin",
            "periodo",
            "receta_id",
            "receta",
            "forecast",
            "real",
            "delta",
            "abs_error",
            "variacion_pct",
            "status",
        ]
    )
    for w in windows:
        for row in w.get("top_errors") or []:
            writer.writerow(
                [
                    w.get("window_start") or "",
                    w.get("window_end") or "",
                    w.get("periodo") or "",
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    f"{Decimal(str(row.get('forecast_qty') or 0)):.3f}",
                    f"{Decimal(str(row.get('actual_qty') or 0)):.3f}",
                    f"{Decimal(str(row.get('delta_qty') or 0)):.3f}",
                    f"{Decimal(str(row.get('abs_error') or 0)):.3f}",
                    f"{Decimal(str(row.get('variacion_pct'))):.1f}" if row.get("variacion_pct") is not None else "",
                    row.get("status") or "",
                ]
            )
    return response


def _forecast_insights_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    scope = payload.get("scope") or {}
    totals = payload.get("totales") or {}
    seasonality = payload.get("seasonality") or {}
    month_rows = seasonality.get("by_month") or []
    weekday_rows = seasonality.get("by_weekday") or []
    top_rows = payload.get("top_recetas") or []

    start_txt = str(scope.get("fecha_desde") or "").replace("-", "")
    end_txt = str(scope.get("fecha_hasta") or "").replace("-", "")
    if not start_txt or not end_txt:
        today_txt = timezone.localdate().strftime("%Y%m%d")
        start_txt = start_txt or today_txt
        end_txt = end_txt or today_txt
    filename = f"api_forecast_insights_{start_txt}_{end_txt}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Meses", int(scope.get("months") or 0)])
        ws_resumen.append(["Fecha desde", scope.get("fecha_desde") or ""])
        ws_resumen.append(["Fecha hasta", scope.get("fecha_hasta") or ""])
        ws_resumen.append(["Sucursal", scope.get("sucursal") or "Todas"])
        ws_resumen.append(["Receta", scope.get("receta") or "Todas"])
        ws_resumen.append(["Top", int(scope.get("top") or 0)])
        ws_resumen.append(["Offset top", int(scope.get("offset_top") or 0)])
        ws_resumen.append(["Filas", int(totals.get("filas") or 0)])
        ws_resumen.append(["Dias con venta", int(totals.get("dias_con_venta") or 0)])
        ws_resumen.append(["Recetas", int(totals.get("recetas") or 0)])
        ws_resumen.append(["Top recetas total", int(totals.get("top_recetas_total") or 0)])
        ws_resumen.append(["Top recetas returned", int(totals.get("top_recetas_returned") or 0)])
        ws_resumen.append(["Cantidad total", float(totals.get("cantidad_total") or 0)])
        ws_resumen.append(["Promedio diario", float(totals.get("promedio_diario") or 0)])

        ws_mes = wb.create_sheet("EstacionalidadMes")
        ws_mes.append(["Mes", "Etiqueta", "Muestras", "Promedio", "Indice %"])
        for row in month_rows:
            ws_mes.append(
                [
                    int(row.get("month") or 0),
                    row.get("label") or "",
                    int(row.get("samples") or 0),
                    float(row.get("avg_qty") or 0),
                    float(row.get("index_pct") or 0),
                ]
            )

        ws_dia = wb.create_sheet("EstacionalidadDia")
        ws_dia.append(["Dia semana", "Etiqueta", "Muestras", "Promedio", "Indice %"])
        for row in weekday_rows:
            ws_dia.append(
                [
                    int(row.get("weekday") or 0),
                    row.get("label") or "",
                    int(row.get("samples") or 0),
                    float(row.get("avg_qty") or 0),
                    float(row.get("index_pct") or 0),
                ]
            )

        ws_top = wb.create_sheet("TopRecetas")
        ws_top.append(["Receta ID", "Receta", "Cantidad total", "Promedio dia activo", "Dias con venta", "Participacion %"])
        for row in top_rows:
            ws_top.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    float(row.get("cantidad_total") or 0),
                    float(row.get("promedio_dia_activo") or 0),
                    int(row.get("dias_con_venta") or 0),
                    float(row.get("participacion_pct") or 0),
                ]
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Meses", int(scope.get("months") or 0)])
    writer.writerow(["Fecha desde", scope.get("fecha_desde") or ""])
    writer.writerow(["Fecha hasta", scope.get("fecha_hasta") or ""])
    writer.writerow(["Sucursal", scope.get("sucursal") or "Todas"])
    writer.writerow(["Receta", scope.get("receta") or "Todas"])
    writer.writerow(["Top", int(scope.get("top") or 0)])
    writer.writerow(["Offset top", int(scope.get("offset_top") or 0)])
    writer.writerow(["Filas", int(totals.get("filas") or 0)])
    writer.writerow(["Dias con venta", int(totals.get("dias_con_venta") or 0)])
    writer.writerow(["Recetas", int(totals.get("recetas") or 0)])
    writer.writerow(["Top recetas total", int(totals.get("top_recetas_total") or 0)])
    writer.writerow(["Top recetas returned", int(totals.get("top_recetas_returned") or 0)])
    writer.writerow(["Cantidad total", f"{Decimal(str(totals.get('cantidad_total') or 0)):.3f}"])
    writer.writerow(["Promedio diario", f"{Decimal(str(totals.get('promedio_diario') or 0)):.3f}"])

    writer.writerow([])
    writer.writerow(["ESTACIONALIDAD_MES"])
    writer.writerow(["month", "label", "samples", "avg_qty", "index_pct"])
    for row in month_rows:
        writer.writerow(
            [
                int(row.get("month") or 0),
                row.get("label") or "",
                int(row.get("samples") or 0),
                f"{Decimal(str(row.get('avg_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('index_pct') or 0)):.1f}",
            ]
        )

    writer.writerow([])
    writer.writerow(["ESTACIONALIDAD_DIA"])
    writer.writerow(["weekday", "label", "samples", "avg_qty", "index_pct"])
    for row in weekday_rows:
        writer.writerow(
            [
                int(row.get("weekday") or 0),
                row.get("label") or "",
                int(row.get("samples") or 0),
                f"{Decimal(str(row.get('avg_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('index_pct') or 0)):.1f}",
            ]
        )

    writer.writerow([])
    writer.writerow(["TOP_RECETAS"])
    writer.writerow(["receta_id", "receta", "cantidad_total", "promedio_dia_activo", "dias_con_venta", "participacion_pct"])
    for row in top_rows:
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                f"{Decimal(str(row.get('cantidad_total') or 0)):.3f}",
                f"{Decimal(str(row.get('promedio_dia_activo') or 0)):.3f}",
                int(row.get("dias_con_venta") or 0),
                f"{Decimal(str(row.get('participacion_pct') or 0)):.1f}",
            ]
        )
    return response


def _ventas_historial_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    filters = payload.get("filters") or {}
    totals = payload.get("totales") or {}
    items = payload.get("items") or []
    by_sucursal = totals.get("by_sucursal") or []
    stamp = str(filters.get("periodo") or timezone.localdate().strftime("%Y-%m")).replace("-", "")
    filename = f"ventas_historial_{stamp}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Periodo", filters.get("periodo") or ""])
        ws_resumen.append(["Desde", filters.get("fecha_desde") or ""])
        ws_resumen.append(["Hasta", filters.get("fecha_hasta") or ""])
        ws_resumen.append(["Offset", int(filters.get("offset") or 0)])
        ws_resumen.append(["Rows", int(totals.get("rows") or 0)])
        ws_resumen.append(["Rows total", int(totals.get("rows_total") or totals.get("rows") or 0)])
        ws_resumen.append(["Rows returned", int(totals.get("rows_returned") or totals.get("rows") or 0)])
        ws_resumen.append(["Cantidad total", float(totals.get("cantidad_total") or 0)])
        ws_resumen.append(["Tickets total", int(totals.get("tickets_total") or 0)])
        ws_resumen.append(["Monto total", float(totals.get("monto_total") or 0)])

        ws_detalle = wb.create_sheet("Detalle")
        ws_detalle.append(
            [
                "ID",
                "Fecha",
                "Receta ID",
                "Receta",
                "Codigo point",
                "Sucursal ID",
                "Sucursal",
                "Sucursal codigo",
                "Cantidad",
                "Tickets",
                "Monto total",
                "Fuente",
                "Actualizado en",
            ]
        )
        for row in items:
            ws_detalle.append(
                [
                    int(row.get("id") or 0),
                    row.get("fecha") or "",
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    row.get("codigo_point") or "",
                    int(row.get("sucursal_id") or 0) if row.get("sucursal_id") else None,
                    row.get("sucursal") or "",
                    row.get("sucursal_codigo") or "",
                    float(row.get("cantidad") or 0),
                    int(row.get("tickets") or 0),
                    float(row.get("monto_total") or 0),
                    row.get("fuente") or "",
                    str(row.get("actualizado_en") or ""),
                ]
            )

        ws_sucursal = wb.create_sheet("BySucursal")
        ws_sucursal.append(["Sucursal", "Cantidad total"])
        for row in by_sucursal:
            ws_sucursal.append([row.get("sucursal") or "", float(row.get("cantidad_total") or 0)])

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Periodo", filters.get("periodo") or ""])
    writer.writerow(["Desde", filters.get("fecha_desde") or ""])
    writer.writerow(["Hasta", filters.get("fecha_hasta") or ""])
    writer.writerow(["Offset", int(filters.get("offset") or 0)])
    writer.writerow(["Rows", int(totals.get("rows") or 0)])
    writer.writerow(["Rows total", int(totals.get("rows_total") or totals.get("rows") or 0)])
    writer.writerow(["Rows returned", int(totals.get("rows_returned") or totals.get("rows") or 0)])
    writer.writerow(["Cantidad total", f"{Decimal(str(totals.get('cantidad_total') or 0)):.3f}"])
    writer.writerow(["Tickets total", int(totals.get("tickets_total") or 0)])
    writer.writerow(["Monto total", f"{Decimal(str(totals.get('monto_total') or 0)):.3f}"])

    writer.writerow([])
    writer.writerow(["BY_SUCURSAL"])
    writer.writerow(["sucursal", "cantidad_total"])
    for row in by_sucursal:
        writer.writerow(
            [
                row.get("sucursal") or "",
                f"{Decimal(str(row.get('cantidad_total') or 0)):.3f}",
            ]
        )

    writer.writerow([])
    writer.writerow(["DETALLE"])
    writer.writerow(
        [
            "id",
            "fecha",
            "receta_id",
            "receta",
            "codigo_point",
            "sucursal_id",
            "sucursal",
            "sucursal_codigo",
            "cantidad",
            "tickets",
            "monto_total",
            "fuente",
            "actualizado_en",
        ]
    )
    for row in items:
        writer.writerow(
            [
                int(row.get("id") or 0),
                row.get("fecha") or "",
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                row.get("codigo_point") or "",
                int(row.get("sucursal_id") or 0) if row.get("sucursal_id") else "",
                row.get("sucursal") or "",
                row.get("sucursal_codigo") or "",
                f"{Decimal(str(row.get('cantidad') or 0)):.3f}",
                int(row.get("tickets") or 0),
                f"{Decimal(str(row.get('monto_total') or 0)):.3f}",
                row.get("fuente") or "",
                str(row.get("actualizado_en") or ""),
            ]
        )
    return response


def _ventas_pronostico_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    filters = payload.get("filters") or {}
    totals = payload.get("totales") or {}
    items = payload.get("items") or []
    stamp = str(filters.get("periodo") or timezone.localdate().strftime("%Y-%m")).replace("-", "")
    filename = f"ventas_pronostico_{stamp}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Periodo", filters.get("periodo") or ""])
        ws_resumen.append(["Periodo desde", filters.get("periodo_desde") or ""])
        ws_resumen.append(["Periodo hasta", filters.get("periodo_hasta") or ""])
        ws_resumen.append(["Offset", int(filters.get("offset") or 0)])
        ws_resumen.append(["Rows", int(totals.get("rows") or 0)])
        ws_resumen.append(["Rows total", int(totals.get("rows_total") or totals.get("rows") or 0)])
        ws_resumen.append(["Rows returned", int(totals.get("rows_returned") or totals.get("rows") or 0)])
        ws_resumen.append(["Cantidad total", float(totals.get("cantidad_total") or 0)])
        ws_resumen.append(["Periodos", int(totals.get("periodos_count") or 0)])

        ws_detalle = wb.create_sheet("Detalle")
        ws_detalle.append(
            [
                "ID",
                "Receta ID",
                "Receta",
                "Codigo point",
                "Periodo",
                "Cantidad",
                "Fuente",
                "Actualizado en",
            ]
        )
        for row in items:
            ws_detalle.append(
                [
                    int(row.get("id") or 0),
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    row.get("codigo_point") or "",
                    row.get("periodo") or "",
                    float(row.get("cantidad") or 0),
                    row.get("fuente") or "",
                    str(row.get("actualizado_en") or ""),
                ]
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Periodo", filters.get("periodo") or ""])
    writer.writerow(["Periodo desde", filters.get("periodo_desde") or ""])
    writer.writerow(["Periodo hasta", filters.get("periodo_hasta") or ""])
    writer.writerow(["Offset", int(filters.get("offset") or 0)])
    writer.writerow(["Rows", int(totals.get("rows") or 0)])
    writer.writerow(["Rows total", int(totals.get("rows_total") or totals.get("rows") or 0)])
    writer.writerow(["Rows returned", int(totals.get("rows_returned") or totals.get("rows") or 0)])
    writer.writerow(["Cantidad total", f"{Decimal(str(totals.get('cantidad_total') or 0)):.3f}"])
    writer.writerow(["Periodos", int(totals.get("periodos_count") or 0)])
    writer.writerow([])
    writer.writerow(["DETALLE"])
    writer.writerow(["id", "receta_id", "receta", "codigo_point", "periodo", "cantidad", "fuente", "actualizado_en"])
    for row in items:
        writer.writerow(
            [
                int(row.get("id") or 0),
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                row.get("codigo_point") or "",
                row.get("periodo") or "",
                f"{Decimal(str(row.get('cantidad') or 0)):.3f}",
                row.get("fuente") or "",
                str(row.get("actualizado_en") or ""),
            ]
        )
    return response


def _ventas_solicitudes_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    filters = payload.get("filters") or {}
    totals = payload.get("totales") or {}
    by_alcance = totals.get("by_alcance") or {}
    forecast_counts = totals.get("forecast_ref_status") or {}
    items = payload.get("items") or []
    stamp = str(filters.get("periodo") or timezone.localdate().strftime("%Y-%m")).replace("-", "")
    filename = f"ventas_solicitudes_{stamp}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Periodo", filters.get("periodo") or ""])
        ws_resumen.append(["Alcance", filters.get("alcance") or ""])
        ws_resumen.append(["Desde", filters.get("fecha_desde") or ""])
        ws_resumen.append(["Hasta", filters.get("fecha_hasta") or ""])
        ws_resumen.append(["Include forecast ref", "SI" if filters.get("include_forecast_ref") else "NO"])
        ws_resumen.append(["Forecast status", filters.get("forecast_status") or ""])
        ws_resumen.append(["Forecast delta min", float(filters.get("forecast_delta_min") or 0)])
        ws_resumen.append(["Offset", int(filters.get("offset") or 0)])
        ws_resumen.append(["Sort by", filters.get("sort_by") or ""])
        ws_resumen.append(["Sort dir", filters.get("sort_dir") or ""])
        ws_resumen.append(["Rows", int(totals.get("rows") or 0)])
        ws_resumen.append(["Rows total", int(totals.get("rows_total") or totals.get("rows") or 0)])
        ws_resumen.append(["Rows returned", int(totals.get("rows_returned") or totals.get("rows") or 0)])
        ws_resumen.append(["Cantidad total", float(totals.get("cantidad_total") or 0)])
        ws_resumen.append(["MES", int(by_alcance.get(SolicitudVenta.ALCANCE_MES) or 0)])
        ws_resumen.append(["SEMANA", int(by_alcance.get(SolicitudVenta.ALCANCE_SEMANA) or 0)])
        ws_resumen.append(["FIN_SEMANA", int(by_alcance.get(SolicitudVenta.ALCANCE_FIN_SEMANA) or 0)])
        ws_resumen.append(["Forecast status SOBRE", int(forecast_counts.get("SOBRE") or 0)])
        ws_resumen.append(["Forecast status BAJO", int(forecast_counts.get("BAJO") or 0)])
        ws_resumen.append(["Forecast status OK", int(forecast_counts.get("OK") or 0)])
        ws_resumen.append(["Forecast status SIN_FORECAST", int(forecast_counts.get("SIN_FORECAST") or 0)])

        ws_detalle = wb.create_sheet("Detalle")
        ws_detalle.append(
            [
                "ID",
                "Receta ID",
                "Receta",
                "Codigo point",
                "Sucursal ID",
                "Sucursal",
                "Sucursal codigo",
                "Alcance",
                "Periodo",
                "Fecha inicio",
                "Fecha fin",
                "Cantidad",
                "Forecast status",
                "Forecast qty",
                "Delta solicitud vs forecast",
                "Fuente",
                "Actualizado en",
            ]
        )
        for row in items:
            forecast_ref = row.get("forecast_ref") or {}
            ws_detalle.append(
                [
                    int(row.get("id") or 0),
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    row.get("codigo_point") or "",
                    int(row.get("sucursal_id") or 0) if row.get("sucursal_id") else None,
                    row.get("sucursal") or "",
                    row.get("sucursal_codigo") or "",
                    row.get("alcance") or "",
                    row.get("periodo") or "",
                    row.get("fecha_inicio") or "",
                    row.get("fecha_fin") or "",
                    float(row.get("cantidad") or 0),
                    forecast_ref.get("status") or "",
                    float(forecast_ref.get("forecast_qty") or 0),
                    float(forecast_ref.get("delta_solicitud_vs_forecast") or 0),
                    row.get("fuente") or "",
                    str(row.get("actualizado_en") or ""),
                ]
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Periodo", filters.get("periodo") or ""])
    writer.writerow(["Alcance", filters.get("alcance") or ""])
    writer.writerow(["Desde", filters.get("fecha_desde") or ""])
    writer.writerow(["Hasta", filters.get("fecha_hasta") or ""])
    writer.writerow(["Include forecast ref", "SI" if filters.get("include_forecast_ref") else "NO"])
    writer.writerow(["Forecast status", filters.get("forecast_status") or ""])
    writer.writerow(["Forecast delta min", f"{Decimal(str(filters.get('forecast_delta_min') or 0)):.3f}"])
    writer.writerow(["Offset", int(filters.get("offset") or 0)])
    writer.writerow(["Sort by", filters.get("sort_by") or ""])
    writer.writerow(["Sort dir", filters.get("sort_dir") or ""])
    writer.writerow(["Rows", int(totals.get("rows") or 0)])
    writer.writerow(["Rows total", int(totals.get("rows_total") or totals.get("rows") or 0)])
    writer.writerow(["Rows returned", int(totals.get("rows_returned") or totals.get("rows") or 0)])
    writer.writerow(["Cantidad total", f"{Decimal(str(totals.get('cantidad_total') or 0)):.3f}"])
    writer.writerow(["MES", int(by_alcance.get(SolicitudVenta.ALCANCE_MES) or 0)])
    writer.writerow(["SEMANA", int(by_alcance.get(SolicitudVenta.ALCANCE_SEMANA) or 0)])
    writer.writerow(["FIN_SEMANA", int(by_alcance.get(SolicitudVenta.ALCANCE_FIN_SEMANA) or 0)])
    writer.writerow(["Forecast status SOBRE", int(forecast_counts.get("SOBRE") or 0)])
    writer.writerow(["Forecast status BAJO", int(forecast_counts.get("BAJO") or 0)])
    writer.writerow(["Forecast status OK", int(forecast_counts.get("OK") or 0)])
    writer.writerow(["Forecast status SIN_FORECAST", int(forecast_counts.get("SIN_FORECAST") or 0)])
    writer.writerow([])
    writer.writerow(["DETALLE"])
    writer.writerow(
        [
            "id",
            "receta_id",
            "receta",
            "codigo_point",
            "sucursal_id",
            "sucursal",
            "sucursal_codigo",
            "alcance",
            "periodo",
            "fecha_inicio",
            "fecha_fin",
            "cantidad",
            "forecast_status",
            "forecast_qty",
            "delta_solicitud_vs_forecast",
            "fuente",
            "actualizado_en",
        ]
    )
    for row in items:
        forecast_ref = row.get("forecast_ref") or {}
        writer.writerow(
            [
                int(row.get("id") or 0),
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                row.get("codigo_point") or "",
                int(row.get("sucursal_id") or 0) if row.get("sucursal_id") else "",
                row.get("sucursal") or "",
                row.get("sucursal_codigo") or "",
                row.get("alcance") or "",
                row.get("periodo") or "",
                row.get("fecha_inicio") or "",
                row.get("fecha_fin") or "",
                f"{Decimal(str(row.get('cantidad') or 0)):.3f}",
                forecast_ref.get("status") or "",
                f"{Decimal(str(forecast_ref.get('forecast_qty') or 0)):.3f}",
                f"{Decimal(str(forecast_ref.get('delta_solicitud_vs_forecast') or 0)):.3f}",
                row.get("fuente") or "",
                str(row.get("actualizado_en") or ""),
            ]
        )
    return response


def _ventas_solicitud_aplicar_forecast_export_response(payload: dict[str, Any], export_format: str) -> HttpResponse:
    scope = payload.get("scope") or {}
    updated = payload.get("updated") or {}
    adjusted_rows = payload.get("adjusted_rows") or []
    compare = payload.get("compare_solicitud") or {}
    compare_totals = compare.get("totals") or {}
    compare_rows = compare.get("rows") or []
    periodo = str(scope.get("periodo") or _normalize_periodo_mes(None)).replace("-", "")
    sucursal_id = int(scope.get("sucursal_id") or 0)
    filename = f"ventas_solicitud_aplicar_forecast_{periodo}_{sucursal_id}.{export_format}"

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Alcance", scope.get("alcance") or ""])
        ws_resumen.append(["Periodo", scope.get("periodo") or ""])
        ws_resumen.append(["Inicio", scope.get("target_start") or ""])
        ws_resumen.append(["Fin", scope.get("target_end") or ""])
        ws_resumen.append(["Sucursal ID", sucursal_id if sucursal_id > 0 else None])
        ws_resumen.append(["Sucursal", scope.get("sucursal_nombre") or ""])
        ws_resumen.append(["Modo", scope.get("modo") or ""])
        ws_resumen.append(["Escenario", scope.get("escenario") or "base"])
        ws_resumen.append(["Confianza minima %", float(scope.get("min_confianza_pct") or 0)])
        ws_resumen.append(["Dry run", "SI" if updated.get("dry_run") else "NO"])
        ws_resumen.append(["Created", int(updated.get("created") or 0)])
        ws_resumen.append(["Updated", int(updated.get("updated") or 0)])
        ws_resumen.append(["Skipped", int(updated.get("skipped") or 0)])
        ws_resumen.append(["Skipped cap", int(updated.get("skipped_cap") or 0)])
        ws_resumen.append(["Applied", int(updated.get("applied") or 0)])
        ws_resumen.append(["Compare forecast total", float(compare_totals.get("forecast_total") or 0)])
        ws_resumen.append(["Compare solicitud total", float(compare_totals.get("solicitud_total") or 0)])
        ws_resumen.append(["Compare delta total", float(compare_totals.get("delta_total") or 0)])
        ws_resumen.append(["Compare en rango", int(compare_totals.get("en_rango_count") or 0)])
        ws_resumen.append(["Compare sobre rango", int(compare_totals.get("sobre_rango_count") or 0)])
        ws_resumen.append(["Compare bajo rango", int(compare_totals.get("bajo_rango_count") or 0)])

        ws_ajustes = wb.create_sheet("Ajustes")
        ws_ajustes.append(
            [
                "Receta ID",
                "Receta",
                "Cantidad anterior",
                "Cantidad nueva",
                "Variacion %",
                "Accion",
                "Status previo",
            ]
        )
        for row in adjusted_rows:
            ws_ajustes.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    float(row.get("anterior") or 0),
                    float(row.get("nueva") or 0),
                    float(row.get("variacion_pct")) if row.get("variacion_pct") is not None else None,
                    row.get("accion") or "",
                    row.get("status_before") or "",
                ]
            )

        ws_compare = wb.create_sheet("Compare")
        ws_compare.append(
            [
                "Receta ID",
                "Receta",
                "Forecast",
                "Forecast base",
                "Forecast bajo",
                "Forecast alto",
                "Solicitud",
                "Delta",
                "Variacion %",
                "Status",
                "Status rango",
            ]
        )
        for row in compare_rows:
            ws_compare.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    float(row.get("forecast_qty") or 0),
                    float(row.get("forecast_base") or 0),
                    float(row.get("forecast_low") or 0),
                    float(row.get("forecast_high") or 0),
                    float(row.get("solicitud_qty") or 0),
                    float(row.get("delta_qty") or 0),
                    float(row.get("variacion_pct")) if row.get("variacion_pct") is not None else None,
                    row.get("status") or "",
                    row.get("status_rango") or "",
                ]
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Alcance", scope.get("alcance") or ""])
    writer.writerow(["Periodo", scope.get("periodo") or ""])
    writer.writerow(["Inicio", scope.get("target_start") or ""])
    writer.writerow(["Fin", scope.get("target_end") or ""])
    writer.writerow(["Sucursal ID", sucursal_id if sucursal_id > 0 else ""])
    writer.writerow(["Sucursal", scope.get("sucursal_nombre") or ""])
    writer.writerow(["Modo", scope.get("modo") or ""])
    writer.writerow(["Escenario", scope.get("escenario") or "base"])
    writer.writerow(["Confianza minima %", f"{Decimal(str(scope.get('min_confianza_pct') or 0)):.1f}"])
    writer.writerow(["Dry run", "SI" if updated.get("dry_run") else "NO"])
    writer.writerow(["Created", int(updated.get("created") or 0)])
    writer.writerow(["Updated", int(updated.get("updated") or 0)])
    writer.writerow(["Skipped", int(updated.get("skipped") or 0)])
    writer.writerow(["Skipped cap", int(updated.get("skipped_cap") or 0)])
    writer.writerow(["Applied", int(updated.get("applied") or 0)])
    writer.writerow(["Compare forecast total", f"{Decimal(str(compare_totals.get('forecast_total') or 0)):.3f}"])
    writer.writerow(["Compare solicitud total", f"{Decimal(str(compare_totals.get('solicitud_total') or 0)):.3f}"])
    writer.writerow(["Compare delta total", f"{Decimal(str(compare_totals.get('delta_total') or 0)):.3f}"])
    writer.writerow(["Compare en rango", int(compare_totals.get("en_rango_count") or 0)])
    writer.writerow(["Compare sobre rango", int(compare_totals.get("sobre_rango_count") or 0)])
    writer.writerow(["Compare bajo rango", int(compare_totals.get("bajo_rango_count") or 0)])
    writer.writerow([])
    writer.writerow(["AJUSTES"])
    writer.writerow(["receta_id", "receta", "anterior", "nueva", "variacion_pct", "accion", "status_before"])
    for row in adjusted_rows:
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                f"{Decimal(str(row.get('anterior') or 0)):.3f}",
                f"{Decimal(str(row.get('nueva') or 0)):.3f}",
                f"{Decimal(str(row.get('variacion_pct'))):.1f}" if row.get("variacion_pct") is not None else "",
                row.get("accion") or "",
                row.get("status_before") or "",
            ]
        )
    writer.writerow([])
    writer.writerow(["COMPARE_SOLICITUD"])
    writer.writerow(
        [
            "receta_id",
            "receta",
            "forecast",
            "forecast_base",
            "forecast_bajo",
            "forecast_alto",
            "solicitud",
            "delta",
            "variacion_pct",
            "status",
            "status_rango",
        ]
    )
    for row in compare_rows:
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                f"{Decimal(str(row.get('forecast_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_base') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_low') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_high') or 0)):.3f}",
                f"{Decimal(str(row.get('solicitud_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('delta_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('variacion_pct'))):.1f}" if row.get("variacion_pct") is not None else "",
                row.get("status") or "",
                row.get("status_rango") or "",
            ]
        )
    return response



class VentaHistoricaListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.has_perm("recetas.view_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para consultar historial de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )
        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        q = (request.GET.get("q") or "").strip()
        periodo = (request.GET.get("periodo") or request.GET.get("mes") or "").strip()
        sucursal_id_raw = (request.GET.get("sucursal_id") or "").strip()
        receta_id_raw = (request.GET.get("receta_id") or "").strip()
        fecha_desde_raw = (request.GET.get("fecha_desde") or "").strip()
        fecha_hasta_raw = (request.GET.get("fecha_hasta") or "").strip()
        limit = _parse_bounded_int(request.GET.get("limit", 150), default=150, min_value=1, max_value=1000)
        offset = _parse_bounded_int(request.GET.get("offset", 0), default=0, min_value=0, max_value=200000)

        fecha_desde = _parse_iso_date(fecha_desde_raw)
        if fecha_desde_raw and fecha_desde is None:
            return Response(
                {"detail": "fecha_desde inválida. Usa formato YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        fecha_hasta = _parse_iso_date(fecha_hasta_raw)
        if fecha_hasta_raw and fecha_hasta is None:
            return Response(
                {"detail": "fecha_hasta inválida. Usa formato YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if fecha_desde and fecha_hasta and fecha_hasta < fecha_desde:
            return Response(
                {"detail": "fecha_hasta no puede ser menor que fecha_desde."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = VentaHistorica.objects.select_related("receta", "sucursal").order_by("-fecha", "-id")
        parsed_period = _parse_period(periodo)
        if periodo and not parsed_period:
            return Response(
                {"detail": "periodo inválido. Usa formato YYYY-MM."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if parsed_period:
            year, month = parsed_period
            qs = qs.filter(fecha__year=year, fecha__month=month)
            periodo = f"{year:04d}-{month:02d}"
        else:
            periodo = ""

        if fecha_desde:
            qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha__lte=fecha_hasta)

        if sucursal_id_raw:
            if not sucursal_id_raw.isdigit():
                return Response(
                    {"detail": "sucursal_id debe ser numérico."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            qs = qs.filter(sucursal_id=int(sucursal_id_raw))

        if receta_id_raw:
            if not receta_id_raw.isdigit():
                return Response(
                    {"detail": "receta_id debe ser numérico."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            qs = qs.filter(receta_id=int(receta_id_raw))

        if q:
            qs = qs.filter(
                Q(receta__nombre__icontains=q)
                | Q(receta__codigo_point__icontains=q)
                | Q(sucursal__nombre__icontains=q)
                | Q(sucursal__codigo__icontains=q)
                | Q(fuente__icontains=q)
            )

        total_rows = qs.count()
        agg = qs.aggregate(
            cantidad_total=Sum("cantidad"),
            tickets_total=Sum("tickets"),
            monto_total=Sum("monto_total"),
        )
        cantidad_total = _to_decimal(agg.get("cantidad_total"))
        tickets_total = int(agg.get("tickets_total") or 0)
        monto_total = _to_decimal(agg.get("monto_total"))
        by_sucursal: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
        branch_totals_qs = qs.order_by().values("sucursal__codigo").annotate(total=Sum("cantidad")).order_by("-total")
        for row in branch_totals_qs:
            sucursal_key = str(row.get("sucursal__codigo") or "GLOBAL")
            by_sucursal[sucursal_key] += _to_decimal(row.get("total"))

        rows = list(qs[offset : offset + limit])
        items = []

        for r in rows:
            cantidad = _to_decimal(r.cantidad)
            monto = _to_decimal(r.monto_total)
            items.append(
                {
                    "id": r.id,
                    "fecha": str(r.fecha),
                    "receta_id": r.receta_id,
                    "receta": r.receta.nombre,
                    "codigo_point": r.receta.codigo_point,
                    "sucursal_id": r.sucursal_id,
                    "sucursal": r.sucursal.nombre if r.sucursal_id and r.sucursal else "",
                    "sucursal_codigo": r.sucursal.codigo if r.sucursal_id and r.sucursal else "",
                    "cantidad": str(cantidad),
                    "tickets": int(r.tickets or 0),
                    "monto_total": str(monto),
                    "fuente": r.fuente,
                    "actualizado_en": r.actualizado_en,
                }
            )

        sucursales_payload = [
            {"sucursal": k, "cantidad_total": str(v)}
            for k, v in sorted(by_sucursal.items(), key=lambda entry: entry[1], reverse=True)
        ]

        payload = {
            "filters": {
                "q": q,
                "periodo": periodo,
                "sucursal_id": sucursal_id_raw,
                "receta_id": receta_id_raw,
                "fecha_desde": str(fecha_desde) if fecha_desde else "",
                "fecha_hasta": str(fecha_hasta) if fecha_hasta else "",
                "limit": limit,
                "offset": offset,
            },
            "totales": {
                "rows": len(items),
                "rows_total": int(total_rows),
                "rows_returned": len(items),
                "cantidad_total": str(cantidad_total),
                "tickets_total": tickets_total,
                "monto_total": str(monto_total),
                "by_sucursal": sucursales_payload,
            },
            "items": items,
        }
        if export_format:
            return _ventas_historial_export_response(payload, export_format)
        return Response(payload, status=status.HTTP_200_OK)


class PronosticoVentaListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.has_perm("recetas.view_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para consultar pronósticos de venta."},
                status=status.HTTP_403_FORBIDDEN,
            )
        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        q = (request.GET.get("q") or "").strip()
        periodo = (request.GET.get("periodo") or "").strip()
        periodo_desde = (request.GET.get("periodo_desde") or "").strip()
        periodo_hasta = (request.GET.get("periodo_hasta") or "").strip()
        receta_id_raw = (request.GET.get("receta_id") or "").strip()
        limit = _parse_bounded_int(request.GET.get("limit", 150), default=150, min_value=1, max_value=1000)
        offset = _parse_bounded_int(request.GET.get("offset", 0), default=0, min_value=0, max_value=200000)

        if periodo:
            parsed_period = _parse_period(periodo)
            if not parsed_period:
                return Response(
                    {"detail": "periodo inválido. Usa formato YYYY-MM."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            year, month = parsed_period
            periodo = f"{year:04d}-{month:02d}"

        if periodo_desde:
            parsed_since = _parse_period(periodo_desde)
            if not parsed_since:
                return Response(
                    {"detail": "periodo_desde inválido. Usa formato YYYY-MM."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            periodo_desde = f"{parsed_since[0]:04d}-{parsed_since[1]:02d}"

        if periodo_hasta:
            parsed_until = _parse_period(periodo_hasta)
            if not parsed_until:
                return Response(
                    {"detail": "periodo_hasta inválido. Usa formato YYYY-MM."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            periodo_hasta = f"{parsed_until[0]:04d}-{parsed_until[1]:02d}"

        if periodo_desde and periodo_hasta and periodo_hasta < periodo_desde:
            return Response(
                {"detail": "periodo_hasta no puede ser menor que periodo_desde."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = PronosticoVenta.objects.select_related("receta").order_by("-periodo", "receta__nombre")
        if periodo:
            qs = qs.filter(periodo=periodo)
        if periodo_desde:
            qs = qs.filter(periodo__gte=periodo_desde)
        if periodo_hasta:
            qs = qs.filter(periodo__lte=periodo_hasta)

        if receta_id_raw:
            if not receta_id_raw.isdigit():
                return Response(
                    {"detail": "receta_id debe ser numérico."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            qs = qs.filter(receta_id=int(receta_id_raw))

        if q:
            qs = qs.filter(
                Q(receta__nombre__icontains=q)
                | Q(receta__codigo_point__icontains=q)
                | Q(fuente__icontains=q)
            )

        total_rows = qs.count()
        cantidad_total = _to_decimal(qs.aggregate(total=Sum("cantidad")).get("total"))
        periodos_count = qs.values("periodo").distinct().count()
        rows = list(qs[offset : offset + limit])
        items = []
        for r in rows:
            qty = _to_decimal(r.cantidad)
            items.append(
                {
                    "id": r.id,
                    "receta_id": r.receta_id,
                    "receta": r.receta.nombre,
                    "codigo_point": r.receta.codigo_point,
                    "periodo": r.periodo,
                    "cantidad": str(qty),
                    "fuente": r.fuente,
                    "actualizado_en": r.actualizado_en,
                }
            )

        payload = {
            "filters": {
                "q": q,
                "periodo": periodo,
                "periodo_desde": periodo_desde,
                "periodo_hasta": periodo_hasta,
                "receta_id": receta_id_raw,
                "limit": limit,
                "offset": offset,
            },
            "totales": {
                "rows": len(items),
                "rows_total": int(total_rows),
                "rows_returned": len(items),
                "cantidad_total": str(cantidad_total),
                "periodos_count": int(periodos_count),
            },
            "items": items,
        }
        if export_format:
            return _ventas_pronostico_export_response(payload, export_format)
        return Response(payload, status=status.HTTP_200_OK)


class VentasPipelineResumenView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.has_perm("recetas.view_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para consultar resumen del pipeline de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )

        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        periodo = (request.GET.get("periodo") or "").strip()
        top = _parse_bounded_int(request.GET.get("top", 120), default=120, min_value=1, max_value=500)
        offset = _parse_bounded_int(request.GET.get("offset", 0), default=0, min_value=0, max_value=5000)
        top_sucursales = _parse_bounded_int(request.GET.get("top_sucursales", 120), default=120, min_value=1, max_value=500)
        offset_sucursales = _parse_bounded_int(
            request.GET.get("offset_sucursales", 0),
            default=0,
            min_value=0,
            max_value=5000,
        )
        status_filter = (request.GET.get("status") or "").strip().upper()
        allowed_status = {"", "SOBRE", "BAJO", "OK", "SIN_SOLICITUD", "SIN_MOV", "DESVIADAS"}
        if status_filter not in allowed_status:
            return Response(
                {"detail": "status inválido. Usa SOBRE, BAJO, OK, SIN_SOLICITUD, SIN_MOV o DESVIADAS."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        delta_min = _to_decimal(request.GET.get("delta_min"), default=Decimal("0"))
        if delta_min < 0:
            delta_min = Decimal("0")
        sort_by = (request.GET.get("sort_by") or "delta_abs").strip().lower()
        sort_dir = (request.GET.get("sort_dir") or "desc").strip().lower()
        sort_sucursales_by = (request.GET.get("sort_sucursales_by") or "delta_abs").strip().lower()
        sort_sucursales_dir = (request.GET.get("sort_sucursales_dir") or "desc").strip().lower()
        q = (request.GET.get("q") or "").strip()

        allowed_sort_rows = {
            "delta_abs",
            "delta",
            "historial",
            "solicitud",
            "pronostico",
            "cobertura_pct",
            "cumplimiento_pct",
            "receta",
        }
        allowed_sort_sucursales = {"delta_abs", "delta", "historial", "solicitud", "cumplimiento_pct", "sucursal"}
        if sort_by not in allowed_sort_rows:
            return Response(
                {
                    "detail": (
                        "sort_by inválido. Usa delta_abs, delta, historial, solicitud, pronostico, cobertura_pct, "
                        "cumplimiento_pct o receta."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        if sort_sucursales_by not in allowed_sort_sucursales:
            return Response(
                {"detail": "sort_sucursales_by inválido. Usa delta_abs, delta, historial, solicitud, cumplimiento_pct o sucursal."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if sort_dir not in {"asc", "desc"}:
            return Response(
                {"detail": "sort_dir inválido. Usa asc o desc."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if sort_sucursales_dir not in {"asc", "desc"}:
            return Response(
                {"detail": "sort_sucursales_dir inválido. Usa asc o desc."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        def _status_match(tag: str) -> bool:
            if not status_filter:
                return True
            if status_filter == "DESVIADAS":
                return tag in {"SOBRE", "BAJO"}
            return tag == status_filter

        if periodo:
            parsed_period = _parse_period(periodo)
            if not parsed_period:
                return Response(
                    {"detail": "periodo inválido. Usa formato YYYY-MM."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            periodo = f"{parsed_period[0]:04d}-{parsed_period[1]:02d}"
        else:
            today = timezone.localdate()
            periodo = f"{today.year:04d}-{today.month:02d}"
            parsed_period = (today.year, today.month)
        year, month = parsed_period

        incluir_preparaciones = _parse_bool(request.GET.get("incluir_preparaciones"), default=False)
        sucursal = None
        sucursal_id_raw = (request.GET.get("sucursal_id") or "").strip()
        if sucursal_id_raw:
            if not sucursal_id_raw.isdigit():
                return Response(
                    {"detail": "sucursal_id debe ser numérico."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            sucursal = Sucursal.objects.filter(pk=int(sucursal_id_raw), activa=True).first()
            if sucursal is None:
                return Response(
                    {"detail": "Sucursal no encontrada o inactiva."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        historial_qs = VentaHistorica.objects.filter(fecha__year=year, fecha__month=month)
        solicitudes_qs = SolicitudVenta.objects.filter(periodo=periodo)
        pronostico_qs = PronosticoVenta.objects.filter(periodo=periodo)

        if q:
            historial_qs = historial_qs.filter(
                Q(receta__nombre__icontains=q)
                | Q(receta__codigo_point__icontains=q)
            )
            solicitudes_qs = solicitudes_qs.filter(
                Q(receta__nombre__icontains=q)
                | Q(receta__codigo_point__icontains=q)
            )
            pronostico_qs = pronostico_qs.filter(
                Q(receta__nombre__icontains=q)
                | Q(receta__codigo_point__icontains=q)
            )

        if sucursal:
            historial_qs = historial_qs.filter(sucursal=sucursal)
            solicitudes_qs = solicitudes_qs.filter(sucursal=sucursal)

        if not incluir_preparaciones:
            historial_qs = historial_qs.filter(receta__tipo=Receta.TIPO_PRODUCTO_FINAL)
            solicitudes_qs = solicitudes_qs.filter(receta__tipo=Receta.TIPO_PRODUCTO_FINAL)
            pronostico_qs = pronostico_qs.filter(receta__tipo=Receta.TIPO_PRODUCTO_FINAL)

        historial_total = _to_decimal(historial_qs.aggregate(total=Sum("cantidad"))["total"])
        solicitud_total = _to_decimal(solicitudes_qs.aggregate(total=Sum("cantidad"))["total"])
        pronostico_total = _to_decimal(pronostico_qs.aggregate(total=Sum("cantidad"))["total"])

        by_alcance = {k: Decimal("0") for k in [SolicitudVenta.ALCANCE_MES, SolicitudVenta.ALCANCE_SEMANA, SolicitudVenta.ALCANCE_FIN_SEMANA]}
        for row in solicitudes_qs.values("alcance").annotate(total=Sum("cantidad")):
            by_alcance[row["alcance"]] = _to_decimal(row["total"])

        historial_sucursal_map = {
            int(row["sucursal_id"]): _to_decimal(row["total"])
            for row in historial_qs.values("sucursal_id").annotate(total=Sum("cantidad"))
            if row.get("sucursal_id") is not None
        }
        solicitud_sucursal_map = {
            int(row["sucursal_id"]): _to_decimal(row["total"])
            for row in solicitudes_qs.values("sucursal_id").annotate(total=Sum("cantidad"))
            if row.get("sucursal_id") is not None
        }
        sucursal_ids = sorted(set(historial_sucursal_map.keys()) | set(solicitud_sucursal_map.keys()))
        sucursal_map = {
            int(sid): (codigo, nombre)
            for sid, codigo, nombre in Sucursal.objects.filter(id__in=sucursal_ids).values_list("id", "codigo", "nombre")
        }
        by_sucursal_rows_tmp: list[dict[str, Any]] = []
        for sid in sucursal_ids:
            hist_qty = historial_sucursal_map.get(sid, Decimal("0"))
            sol_qty = solicitud_sucursal_map.get(sid, Decimal("0"))
            delta = hist_qty - sol_qty
            cumplimiento_pct = None
            if sol_qty > 0:
                cumplimiento_pct = ((hist_qty / sol_qty) * Decimal("100")).quantize(Decimal("0.1"))
            if sol_qty > 0:
                threshold = max(Decimal("1"), sol_qty * Decimal("0.10"))
                if delta > threshold:
                    status_tag = "SOBRE"
                elif delta < (threshold * Decimal("-1")):
                    status_tag = "BAJO"
                else:
                    status_tag = "OK"
            elif hist_qty > 0:
                status_tag = "SIN_SOLICITUD"
            else:
                status_tag = "SIN_MOV"

            suc_codigo, suc_nombre = sucursal_map.get(sid, ("", f"Sucursal {sid}"))
            delta_abs = abs(delta)
            by_sucursal_rows_tmp.append(
                {
                    "_delta_abs": _to_float(delta_abs),
                    "sucursal_id": sid,
                    "sucursal_codigo": suc_codigo,
                    "sucursal": suc_nombre,
                    "historial_qty": _to_float(hist_qty),
                    "solicitud_qty": _to_float(sol_qty),
                    "delta_historial_vs_solicitud": _to_float(delta.quantize(Decimal("0.001"))),
                    "cumplimiento_pct": _to_float(cumplimiento_pct) if cumplimiento_pct is not None else None,
                    "status": status_tag,
                }
            )
        by_sucursal_rows_tmp = [
            row
            for row in by_sucursal_rows_tmp
            if _status_match(str(row.get("status") or "")) and Decimal(str(row["_delta_abs"])) >= delta_min
        ]
        by_sucursal_sort_field = {
            "delta_abs": "_delta_abs",
            "delta": "delta_historial_vs_solicitud",
            "historial": "historial_qty",
            "solicitud": "solicitud_qty",
            "cumplimiento_pct": "cumplimiento_pct",
            "sucursal": "sucursal",
        }[sort_sucursales_by]
        for row in by_sucursal_rows_tmp:
            sort_value = row.get(by_sucursal_sort_field)
            if isinstance(sort_value, str):
                row["_sort"] = sort_value.lower()
            elif sort_value is None:
                row["_sort"] = float("-inf")
            else:
                row["_sort"] = _to_float(sort_value)
        by_sucursal_rows_tmp.sort(key=lambda row: row["_sort"], reverse=(sort_sucursales_dir == "desc"))
        by_sucursal_filtered = len(by_sucursal_rows_tmp)
        by_sucursal_rows = []
        for row in by_sucursal_rows_tmp[offset_sucursales : offset_sucursales + top_sucursales]:
            row.pop("_sort", None)
            row.pop("_delta_abs", None)
            by_sucursal_rows.append(row)
        by_sucursal_status_counts = {
            "SOBRE": sum(1 for row in by_sucursal_rows_tmp if row.get("status") == "SOBRE"),
            "BAJO": sum(1 for row in by_sucursal_rows_tmp if row.get("status") == "BAJO"),
            "OK": sum(1 for row in by_sucursal_rows_tmp if row.get("status") == "OK"),
            "SIN_SOLICITUD": sum(1 for row in by_sucursal_rows_tmp if row.get("status") == "SIN_SOLICITUD"),
            "SIN_MOV": sum(1 for row in by_sucursal_rows_tmp if row.get("status") == "SIN_MOV"),
        }

        cobertura_solicitud_pct = None
        if pronostico_total > 0:
            cobertura_solicitud_pct = ((solicitud_total / pronostico_total) * Decimal("100")).quantize(Decimal("0.1"))

        cumplimiento_historial_pct = None
        if solicitud_total > 0:
            cumplimiento_historial_pct = ((historial_total / solicitud_total) * Decimal("100")).quantize(Decimal("0.1"))

        latest_historial = historial_qs.order_by("-actualizado_en").values_list("actualizado_en", flat=True).first()
        latest_pronostico = pronostico_qs.order_by("-actualizado_en").values_list("actualizado_en", flat=True).first()
        latest_solicitud = solicitudes_qs.order_by("-actualizado_en").values_list("actualizado_en", flat=True).first()

        historial_map = {
            int(row["receta_id"]): _to_decimal(row["total"])
            for row in historial_qs.values("receta_id").annotate(total=Sum("cantidad"))
        }
        pronostico_map = {
            int(row["receta_id"]): _to_decimal(row["total"])
            for row in pronostico_qs.values("receta_id").annotate(total=Sum("cantidad"))
        }
        solicitud_map = {
            int(row["receta_id"]): _to_decimal(row["total"])
            for row in solicitudes_qs.values("receta_id").annotate(total=Sum("cantidad"))
        }
        receta_ids = sorted(set(historial_map.keys()) | set(pronostico_map.keys()) | set(solicitud_map.keys()))
        receta_names = {
            int(receta_id): nombre
            for receta_id, nombre in Receta.objects.filter(id__in=receta_ids).values_list("id", "nombre")
        }

        rows_tmp: list[dict[str, Any]] = []
        for receta_id in receta_ids:
            historial_qty = historial_map.get(receta_id, Decimal("0"))
            pronostico_qty = pronostico_map.get(receta_id, Decimal("0"))
            solicitud_qty = solicitud_map.get(receta_id, Decimal("0"))
            delta_solicitud_vs_pronostico = solicitud_qty - pronostico_qty
            delta_historial_vs_solicitud = historial_qty - solicitud_qty

            cobertura_pct = None
            if pronostico_qty > 0:
                cobertura_pct = ((solicitud_qty / pronostico_qty) * Decimal("100")).quantize(Decimal("0.1"))

            cumplimiento_pct = None
            if solicitud_qty > 0:
                cumplimiento_pct = ((historial_qty / solicitud_qty) * Decimal("100")).quantize(Decimal("0.1"))

            status_tag = "SIN_MOV"
            if solicitud_qty > 0:
                threshold = max(Decimal("1"), solicitud_qty * Decimal("0.10"))
                if delta_historial_vs_solicitud > threshold:
                    status_tag = "SOBRE"
                elif delta_historial_vs_solicitud < (threshold * Decimal("-1")):
                    status_tag = "BAJO"
                else:
                    status_tag = "OK"
            elif historial_qty > 0:
                status_tag = "SIN_SOLICITUD"
            elif pronostico_qty > 0:
                status_tag = "SIN_SOLICITUD"

            rows_tmp.append(
                {
                    "_delta_abs": _to_float(abs(delta_historial_vs_solicitud)),
                    "receta_id": receta_id,
                    "receta": receta_names.get(receta_id) or f"Receta {receta_id}",
                    "historial_qty": _to_float(historial_qty),
                    "pronostico_qty": _to_float(pronostico_qty),
                    "solicitud_qty": _to_float(solicitud_qty),
                    "delta_solicitud_vs_pronostico": _to_float(delta_solicitud_vs_pronostico.quantize(Decimal("0.001"))),
                    "delta_historial_vs_solicitud": _to_float(delta_historial_vs_solicitud.quantize(Decimal("0.001"))),
                    "cobertura_pct": _to_float(cobertura_pct) if cobertura_pct is not None else None,
                    "cumplimiento_pct": _to_float(cumplimiento_pct) if cumplimiento_pct is not None else None,
                    "status": status_tag,
                }
            )
        rows_tmp = [
            row
            for row in rows_tmp
            if _status_match(str(row.get("status") or "")) and Decimal(str(row["_delta_abs"])) >= delta_min
        ]
        row_sort_field = {
            "delta_abs": "_delta_abs",
            "delta": "delta_historial_vs_solicitud",
            "historial": "historial_qty",
            "solicitud": "solicitud_qty",
            "pronostico": "pronostico_qty",
            "cobertura_pct": "cobertura_pct",
            "cumplimiento_pct": "cumplimiento_pct",
            "receta": "receta",
        }[sort_by]
        for row in rows_tmp:
            sort_value = row.get(row_sort_field)
            if isinstance(sort_value, str):
                row["_sort"] = sort_value.lower()
            elif sort_value is None:
                row["_sort"] = float("-inf")
            else:
                row["_sort"] = _to_float(sort_value)
        rows_tmp.sort(key=lambda row: row["_sort"], reverse=(sort_dir == "desc"))
        rows_filtered = len(rows_tmp)
        rows = []
        for row in rows_tmp[offset : offset + top]:
            row.pop("_sort", None)
            row.pop("_delta_abs", None)
            rows.append(row)
        rows_status_counts = {
            "SOBRE": sum(1 for row in rows_tmp if row.get("status") == "SOBRE"),
            "BAJO": sum(1 for row in rows_tmp if row.get("status") == "BAJO"),
            "OK": sum(1 for row in rows_tmp if row.get("status") == "OK"),
            "SIN_SOLICITUD": sum(1 for row in rows_tmp if row.get("status") == "SIN_SOLICITUD"),
            "SIN_MOV": sum(1 for row in rows_tmp if row.get("status") == "SIN_MOV"),
        }

        payload = {
            "scope": {
                "periodo": periodo,
                "sucursal_id": sucursal.id if sucursal else None,
                "sucursal": sucursal.nombre if sucursal else "Todas",
                "incluir_preparaciones": incluir_preparaciones,
                "top": top,
                "offset": offset,
                "top_sucursales": top_sucursales,
                "offset_sucursales": offset_sucursales,
                "sort_by": sort_by,
                "sort_dir": sort_dir,
                "sort_sucursales_by": sort_sucursales_by,
                "sort_sucursales_dir": sort_sucursales_dir,
                "status": status_filter or "",
                "delta_min": _to_float(delta_min),
                "q": q,
            },
            "totales": {
                "historial_qty": _to_float(historial_total),
                "pronostico_qty": _to_float(pronostico_total),
                "solicitud_qty": _to_float(solicitud_total),
                "delta_solicitud_vs_pronostico": _to_float((solicitud_total - pronostico_total).quantize(Decimal("0.001"))),
                "delta_historial_vs_solicitud": _to_float((historial_total - solicitud_total).quantize(Decimal("0.001"))),
                "cobertura_solicitud_pct": _to_float(cobertura_solicitud_pct) if cobertura_solicitud_pct is not None else None,
                "cumplimiento_historial_pct": _to_float(cumplimiento_historial_pct) if cumplimiento_historial_pct is not None else None,
                "historial_recetas": historial_qs.values("receta_id").distinct().count(),
                "pronostico_recetas": pronostico_qs.values("receta_id").distinct().count(),
                "solicitud_recetas": solicitudes_qs.values("receta_id").distinct().count(),
                "rows_count": len(receta_ids),
                "rows_filtered": rows_filtered,
                "rows_returned": len(rows),
                "by_sucursal_filtered": by_sucursal_filtered,
                "by_sucursal_returned": len(by_sucursal_rows),
                "rows_status": rows_status_counts,
                "by_sucursal_status": by_sucursal_status_counts,
            },
            "solicitud_by_alcance": {
                "MES": _to_float(by_alcance[SolicitudVenta.ALCANCE_MES]),
                "SEMANA": _to_float(by_alcance[SolicitudVenta.ALCANCE_SEMANA]),
                "FIN_SEMANA": _to_float(by_alcance[SolicitudVenta.ALCANCE_FIN_SEMANA]),
            },
            "by_sucursal": by_sucursal_rows,
            "latest_updates": {
                "historial": latest_historial,
                "pronostico": latest_pronostico,
                "solicitud": latest_solicitud,
            },
            "rows": rows,
        }
        if export_format:
            return _ventas_pipeline_export_response(payload, export_format)
        return Response(payload, status=status.HTTP_200_OK)


class SolicitudVentaListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.has_perm("recetas.view_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para consultar solicitudes de venta."},
                status=status.HTTP_403_FORBIDDEN,
            )
        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        q = (request.GET.get("q") or "").strip()
        periodo = (request.GET.get("periodo") or "").strip()
        alcance = (request.GET.get("alcance") or "").strip().upper()
        sucursal_id_raw = (request.GET.get("sucursal_id") or "").strip()
        receta_id_raw = (request.GET.get("receta_id") or "").strip()
        fecha_desde_raw = (request.GET.get("fecha_desde") or "").strip()
        fecha_hasta_raw = (request.GET.get("fecha_hasta") or "").strip()
        limit = _parse_bounded_int(request.GET.get("limit", 150), default=150, min_value=1, max_value=1000)
        offset = _parse_bounded_int(request.GET.get("offset", 0), default=0, min_value=0, max_value=200000)
        include_forecast_ref = _parse_bool(request.GET.get("include_forecast_ref"), default=False)
        forecast_status_filter = (request.GET.get("forecast_status") or "").strip().upper()
        forecast_delta_min = _to_decimal(request.GET.get("forecast_delta_min"), default=Decimal("0"))
        sort_by = (request.GET.get("sort_by") or "fecha_inicio").strip().lower()
        sort_dir = (request.GET.get("sort_dir") or "desc").strip().lower()
        if forecast_delta_min < 0:
            forecast_delta_min = Decimal("0")
        allowed_forecast_status = {"", "SOBRE", "BAJO", "OK", "SIN_FORECAST", "DESVIADAS"}
        if forecast_status_filter not in allowed_forecast_status:
            return Response(
                {"detail": "forecast_status inválido. Usa SOBRE, BAJO, OK, SIN_FORECAST o DESVIADAS."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        allowed_sort_by = {
            "fecha_inicio",
            "fecha_fin",
            "cantidad",
            "receta",
            "sucursal",
            "alcance",
            "periodo",
            "forecast_delta",
            "forecast_status",
        }
        if sort_by not in allowed_sort_by:
            return Response(
                {
                    "detail": (
                        "sort_by inválido. Usa fecha_inicio, fecha_fin, cantidad, receta, sucursal, alcance, periodo, "
                        "forecast_delta o forecast_status."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        if sort_dir not in {"asc", "desc"}:
            return Response(
                {"detail": "sort_dir inválido. Usa asc o desc."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if forecast_status_filter or forecast_delta_min > 0:
            include_forecast_ref = True
        if sort_by in {"forecast_delta", "forecast_status"}:
            include_forecast_ref = True

        if periodo:
            parsed_period = _parse_period(periodo)
            if not parsed_period:
                return Response(
                    {"detail": "periodo inválido. Usa formato YYYY-MM."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            periodo = f"{parsed_period[0]:04d}-{parsed_period[1]:02d}"

        allowed_alcance = {
            SolicitudVenta.ALCANCE_MES,
            SolicitudVenta.ALCANCE_SEMANA,
            SolicitudVenta.ALCANCE_FIN_SEMANA,
        }
        if alcance and alcance not in allowed_alcance:
            return Response(
                {"detail": "alcance inválido. Usa MES, SEMANA o FIN_SEMANA."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        fecha_desde = _parse_iso_date(fecha_desde_raw)
        if fecha_desde_raw and fecha_desde is None:
            return Response(
                {"detail": "fecha_desde inválida. Usa formato YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        fecha_hasta = _parse_iso_date(fecha_hasta_raw)
        if fecha_hasta_raw and fecha_hasta is None:
            return Response(
                {"detail": "fecha_hasta inválida. Usa formato YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if fecha_desde and fecha_hasta and fecha_hasta < fecha_desde:
            return Response(
                {"detail": "fecha_hasta no puede ser menor que fecha_desde."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = SolicitudVenta.objects.select_related("receta", "sucursal").order_by("-fecha_inicio", "receta__nombre")
        if periodo:
            qs = qs.filter(periodo=periodo)
        if alcance:
            qs = qs.filter(alcance=alcance)
        if fecha_desde:
            qs = qs.filter(fecha_inicio__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha_fin__lte=fecha_hasta)

        if sucursal_id_raw:
            if not sucursal_id_raw.isdigit():
                return Response(
                    {"detail": "sucursal_id debe ser numérico."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            qs = qs.filter(sucursal_id=int(sucursal_id_raw))

        if receta_id_raw:
            if not receta_id_raw.isdigit():
                return Response(
                    {"detail": "receta_id debe ser numérico."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            qs = qs.filter(receta_id=int(receta_id_raw))

        if q:
            qs = qs.filter(
                Q(receta__nombre__icontains=q)
                | Q(receta__codigo_point__icontains=q)
                | Q(sucursal__nombre__icontains=q)
                | Q(sucursal__codigo__icontains=q)
                | Q(fuente__icontains=q)
            )

        rows = list(qs)
        pronostico_map: dict[tuple[int, str], Decimal] = {}
        if include_forecast_ref and rows:
            receta_ids = {int(r.receta_id) for r in rows}
            periodos = {str(r.periodo) for r in rows if r.periodo}
            if receta_ids and periodos:
                for p in (
                    PronosticoVenta.objects.filter(receta_id__in=receta_ids, periodo__in=periodos)
                    .values("receta_id", "periodo")
                    .annotate(total=Sum("cantidad"))
                ):
                    pronostico_map[(int(p["receta_id"]), str(p["periodo"]))] = _to_decimal(p["total"])

        items = []
        cantidad_total = Decimal("0")
        by_alcance = {k: 0 for k in allowed_alcance}
        forecast_status_counts = {"SOBRE": 0, "BAJO": 0, "OK": 0, "SIN_FORECAST": 0}

        def _forecast_match(status_value: str, delta_value: Decimal) -> bool:
            if forecast_status_filter:
                if forecast_status_filter == "DESVIADAS":
                    if status_value not in {"SOBRE", "BAJO"}:
                        return False
                elif status_value != forecast_status_filter:
                    return False
            if forecast_delta_min > 0 and abs(delta_value) < forecast_delta_min:
                return False
            return True

        for r in rows:
            qty = _to_decimal(r.cantidad)
            item = {
                "id": r.id,
                "receta_id": r.receta_id,
                "receta": r.receta.nombre,
                "codigo_point": r.receta.codigo_point,
                "sucursal_id": r.sucursal_id,
                "sucursal": r.sucursal.nombre if r.sucursal_id and r.sucursal else "",
                "sucursal_codigo": r.sucursal.codigo if r.sucursal_id and r.sucursal else "",
                "alcance": r.alcance,
                "periodo": r.periodo,
                "fecha_inicio": str(r.fecha_inicio),
                "fecha_fin": str(r.fecha_fin),
                "cantidad": str(qty),
                "fuente": r.fuente,
                "actualizado_en": r.actualizado_en,
            }
            if include_forecast_ref:
                forecast_qty = _to_decimal(pronostico_map.get((int(r.receta_id), str(r.periodo))), default=Decimal("0"))
                if forecast_qty > 0:
                    delta_forecast = qty - forecast_qty
                    threshold = max(Decimal("1"), forecast_qty * Decimal("0.10"))
                    if delta_forecast > threshold:
                        forecast_status = "SOBRE"
                    elif delta_forecast < (threshold * Decimal("-1")):
                        forecast_status = "BAJO"
                    else:
                        forecast_status = "OK"
                else:
                    delta_forecast = qty
                    forecast_status = "SIN_FORECAST"

                item["forecast_ref"] = {
                    "status": forecast_status,
                    "forecast_qty": _to_float(forecast_qty),
                    "solicitud_qty": _to_float(qty),
                    "delta_solicitud_vs_forecast": _to_float(delta_forecast.quantize(Decimal("0.001"))),
                }
                if not _forecast_match(forecast_status, delta_forecast):
                    continue
                forecast_status_counts[forecast_status] = forecast_status_counts.get(forecast_status, 0) + 1
            cantidad_total += qty
            by_alcance[r.alcance] = by_alcance.get(r.alcance, 0) + 1
            items.append(item)

        def _sort_value(item: dict[str, Any]) -> Any:
            if sort_by == "cantidad":
                return _to_float(_to_decimal(item.get("cantidad"), default=Decimal("0")))
            if sort_by == "fecha_inicio":
                return _parse_iso_date(str(item.get("fecha_inicio") or "")) or date.min
            if sort_by == "fecha_fin":
                return _parse_iso_date(str(item.get("fecha_fin") or "")) or date.min
            if sort_by == "receta":
                return str(item.get("receta") or "").lower()
            if sort_by == "sucursal":
                return str(item.get("sucursal") or "").lower()
            if sort_by == "alcance":
                return str(item.get("alcance") or "")
            if sort_by == "periodo":
                return str(item.get("periodo") or "")
            if sort_by == "forecast_delta":
                return _to_float((item.get("forecast_ref") or {}).get("delta_solicitud_vs_forecast") or 0)
            if sort_by == "forecast_status":
                status_rank = {"SOBRE": 4, "BAJO": 3, "OK": 2, "SIN_FORECAST": 1}
                return int(status_rank.get(str((item.get("forecast_ref") or {}).get("status") or ""), 0))
            return str(item.get("fecha_inicio") or "")

        items.sort(key=_sort_value, reverse=(sort_dir == "desc"))
        rows_total = len(items)
        items = items[offset : offset + limit]

        payload = {
            "filters": {
                "q": q,
                "periodo": periodo,
                "alcance": alcance,
                "sucursal_id": sucursal_id_raw,
                "receta_id": receta_id_raw,
                "fecha_desde": str(fecha_desde) if fecha_desde else "",
                "fecha_hasta": str(fecha_hasta) if fecha_hasta else "",
                "limit": limit,
                "offset": offset,
                "include_forecast_ref": include_forecast_ref,
                "forecast_status": forecast_status_filter or "",
                "forecast_delta_min": _to_float(forecast_delta_min),
                "sort_by": sort_by,
                "sort_dir": sort_dir,
            },
            "totales": {
                "rows": len(items),
                "rows_total": rows_total,
                "rows_returned": len(items),
                "cantidad_total": str(cantidad_total),
                "by_alcance": by_alcance,
                "forecast_ref_status": forecast_status_counts if include_forecast_ref else {},
            },
            "items": items,
        }
        if export_format:
            return _ventas_solicitudes_export_response(payload, export_format)
        return Response(payload, status=status.HTTP_200_OK)


def _process_pronostico_venta_bulk(
    data: dict[str, Any],
    *,
    dry_run_override: bool | None = None,
) -> dict[str, Any]:
    rows = data["rows"]
    modo = data.get("modo") or "replace"
    fuente = (data.get("fuente") or "API_PRON_BULK").strip()[:40] or "API_PRON_BULK"
    dry_run = bool(data.get("dry_run", True) if dry_run_override is None else dry_run_override)
    stop_on_error = bool(data.get("stop_on_error", False))
    top = int(data.get("top") or 120)

    receta_cache: dict[tuple[int, str, str], Receta | None] = {}
    created = 0
    updated = 0
    skipped = 0
    terminated_early = False
    result_rows: list[dict] = []

    tx_cm = nullcontext() if dry_run else transaction.atomic()
    with tx_cm:
        for index, row in enumerate(rows, start=1):
            receta_id = row.get("receta_id")
            receta_name = str(row.get("receta") or "").strip()
            codigo_point = str(row.get("codigo_point") or "").strip()
            receta = _resolve_receta_bulk_ref(
                receta_id=receta_id,
                receta_name=receta_name,
                codigo_point=codigo_point,
                cache=receta_cache,
            )
            if receta is None:
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "receta_not_found",
                        "receta_id": int(receta_id or 0) or None,
                        "receta_input": receta_name or codigo_point,
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            periodo = _normalize_periodo_mes(row.get("periodo"))
            cantidad = _to_decimal(row.get("cantidad"), default=Decimal("0"))
            existing = PronosticoVenta.objects.filter(receta=receta, periodo=periodo).first()

            previous_qty = _to_decimal(existing.cantidad, default=Decimal("0")) if existing else Decimal("0")
            if existing:
                new_qty = previous_qty + cantidad if modo == "accumulate" else cantidad
                action = "UPDATED"
                updated += 1
            else:
                new_qty = cantidad
                action = "CREATED"
                created += 1

            if not dry_run:
                if existing:
                    existing.cantidad = new_qty
                    existing.fuente = fuente
                    existing.save(update_fields=["cantidad", "fuente", "actualizado_en"])
                else:
                    PronosticoVenta.objects.create(
                        receta=receta,
                        periodo=periodo,
                        cantidad=new_qty,
                        fuente=fuente,
                    )

            result_rows.append(
                {
                    "row": index,
                    "status": action,
                    "receta_id": receta.id,
                    "receta": receta.nombre,
                    "periodo": periodo,
                    "cantidad_prev": float(previous_qty),
                    "cantidad_nueva": float(new_qty),
                }
            )

    error_count = sum(1 for row in result_rows if row.get("status") == "ERROR")
    return {
        "dry_run": dry_run,
        "mode": modo,
        "fuente": fuente,
        "terminated_early": terminated_early,
        "summary": {
            "total_rows": len(rows),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": error_count,
            "applied": 0 if dry_run else (created + updated),
        },
        "rows": result_rows[:top],
    }


class PronosticoVentaImportPreviewView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_pronosticoventa"):
            return Response(
                {"detail": "No tienes permisos para previsualizar pronósticos de venta."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = PronosticoVentaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        payload = _process_pronostico_venta_bulk(ser.validated_data, dry_run_override=True)
        payload["preview"] = True
        return Response(payload, status=status.HTTP_200_OK)


class PronosticoVentaImportConfirmView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_pronosticoventa"):
            return Response(
                {"detail": "No tienes permisos para confirmar importación de pronósticos de venta."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = PronosticoVentaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        payload = _process_pronostico_venta_bulk(ser.validated_data, dry_run_override=False)
        payload["preview"] = False
        return Response(payload, status=status.HTTP_200_OK)


class PronosticoVentaBulkUpsertView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_pronosticoventa"):
            return Response(
                {"detail": "No tienes permisos para importar pronósticos de venta."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = PronosticoVentaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        payload = _process_pronostico_venta_bulk(ser.validated_data)
        return Response(payload, status=status.HTTP_200_OK)


def _process_venta_historica_bulk(
    data: dict[str, Any],
    *,
    dry_run_override: bool | None = None,
) -> dict[str, Any]:
    rows = data["rows"]
    modo = data.get("modo") or "replace"
    fuente = (data.get("fuente") or "API_VENTAS_BULK").strip()[:40] or "API_VENTAS_BULK"
    dry_run = bool(data.get("dry_run", True) if dry_run_override is None else dry_run_override)
    stop_on_error = bool(data.get("stop_on_error", False))
    top = int(data.get("top") or 120)

    default_sucursal = None
    default_sucursal_id = data.get("sucursal_default_id")
    if default_sucursal_id is not None:
        default_sucursal = Sucursal.objects.filter(pk=default_sucursal_id, activa=True).first()
        if default_sucursal is None:
            raise ValueError("Sucursal default no encontrada o inactiva.")

    receta_cache: dict[tuple[int, str, str], Receta | None] = {}
    sucursal_cache: dict[tuple[int, str, str, int], Sucursal | None] = {}
    created = 0
    updated = 0
    skipped = 0
    terminated_early = False
    result_rows: list[dict] = []

    tx_cm = nullcontext() if dry_run else transaction.atomic()
    with tx_cm:
        for index, row in enumerate(rows, start=1):
            receta_id = row.get("receta_id")
            receta_name = str(row.get("receta") or "").strip()
            codigo_point = str(row.get("codigo_point") or "").strip()
            receta = _resolve_receta_bulk_ref(
                receta_id=receta_id,
                receta_name=receta_name,
                codigo_point=codigo_point,
                cache=receta_cache,
            )
            if receta is None:
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "receta_not_found",
                        "receta_id": int(receta_id or 0) or None,
                        "receta_input": receta_name or codigo_point,
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            sucursal_id = row.get("sucursal_id")
            sucursal_name = str(row.get("sucursal") or "").strip()
            sucursal_codigo = str(row.get("sucursal_codigo") or "").strip()
            has_sucursal_ref = bool(sucursal_id) or bool(sucursal_name) or bool(sucursal_codigo)
            sucursal = _resolve_sucursal_bulk_ref(
                sucursal_id=sucursal_id,
                sucursal_name=sucursal_name,
                sucursal_codigo=sucursal_codigo,
                default_sucursal=default_sucursal,
                cache=sucursal_cache,
            )
            if has_sucursal_ref and sucursal is None:
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "sucursal_not_found",
                        "receta_id": receta.id,
                        "receta": receta.nombre,
                        "sucursal_input": sucursal_codigo or sucursal_name or str(sucursal_id),
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            fecha = row["fecha"]
            cantidad = _to_decimal(row.get("cantidad"), default=Decimal("0"))
            tickets = int(row.get("tickets") or 0)
            monto_total_raw = row.get("monto_total", None)
            monto_total = _to_decimal(monto_total_raw) if monto_total_raw is not None else None
            if monto_total is not None and monto_total <= 0:
                monto_total = None

            existing_qs = VentaHistorica.objects.filter(receta=receta, fecha=fecha)
            if sucursal:
                existing_qs = existing_qs.filter(sucursal=sucursal)
            else:
                existing_qs = existing_qs.filter(sucursal__isnull=True)
            existing = existing_qs.order_by("id").first()

            previous_qty = _to_decimal(existing.cantidad, default=Decimal("0")) if existing else Decimal("0")
            previous_tickets = int(existing.tickets or 0) if existing else 0
            if existing:
                if modo == "accumulate":
                    new_qty = previous_qty + cantidad
                    new_tickets = previous_tickets + tickets
                else:
                    new_qty = cantidad
                    new_tickets = tickets
                action = "UPDATED"
                updated += 1
            else:
                new_qty = cantidad
                new_tickets = tickets
                action = "CREATED"
                created += 1

            if not dry_run:
                if existing:
                    existing.cantidad = new_qty
                    existing.tickets = new_tickets
                    if modo == "accumulate":
                        if monto_total is not None:
                            existing.monto_total = _to_decimal(existing.monto_total, default=Decimal("0")) + monto_total
                    else:
                        existing.monto_total = monto_total
                    existing.fuente = fuente
                    existing.save(update_fields=["cantidad", "tickets", "monto_total", "fuente", "actualizado_en"])
                else:
                    VentaHistorica.objects.create(
                        receta=receta,
                        sucursal=sucursal,
                        fecha=fecha,
                        cantidad=new_qty,
                        tickets=new_tickets,
                        monto_total=monto_total,
                        fuente=fuente,
                    )

            result_rows.append(
                {
                    "row": index,
                    "status": action,
                    "receta_id": receta.id,
                    "receta": receta.nombre,
                    "sucursal_id": sucursal.id if sucursal else None,
                    "sucursal": sucursal.nombre if sucursal else "",
                    "fecha": str(fecha),
                    "cantidad_prev": float(previous_qty),
                    "cantidad_nueva": float(new_qty),
                    "tickets_prev": previous_tickets,
                    "tickets_nuevos": new_tickets,
                }
            )

    error_count = sum(1 for row in result_rows if row.get("status") == "ERROR")
    return {
        "dry_run": dry_run,
        "mode": modo,
        "fuente": fuente,
        "terminated_early": terminated_early,
        "summary": {
            "total_rows": len(rows),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": error_count,
            "applied": 0 if dry_run else (created + updated),
        },
        "rows": result_rows[:top],
    }


class VentaHistoricaImportPreviewView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para previsualizar historial de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = VentaHistoricaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_venta_historica_bulk(ser.validated_data, dry_run_override=True)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        payload["preview"] = True
        return Response(payload, status=status.HTTP_200_OK)


class VentaHistoricaImportConfirmView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para confirmar importación de historial de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = VentaHistoricaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_venta_historica_bulk(ser.validated_data, dry_run_override=False)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        payload["preview"] = False
        return Response(payload, status=status.HTTP_200_OK)


class VentaHistoricaBulkUpsertView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para importar historial de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = VentaHistoricaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_venta_historica_bulk(ser.validated_data)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(payload, status=status.HTTP_200_OK)


def _process_solicitud_venta_bulk(
    data: dict[str, Any],
    *,
    dry_run_override: bool | None = None,
) -> dict[str, Any]:
    rows = data["rows"]
    modo = data.get("modo") or "replace"
    fuente = (data.get("fuente") or "API_SOL_BULK").strip()[:40] or "API_SOL_BULK"
    dry_run = bool(data.get("dry_run", True) if dry_run_override is None else dry_run_override)
    stop_on_error = bool(data.get("stop_on_error", False))
    top = int(data.get("top") or 120)

    default_sucursal = None
    default_sucursal_id = data.get("sucursal_default_id")
    if default_sucursal_id is not None:
        default_sucursal = Sucursal.objects.filter(pk=default_sucursal_id, activa=True).first()
        if default_sucursal is None:
            raise ValueError("Sucursal default no encontrada o inactiva.")

    receta_cache: dict[tuple[int, str, str], Receta | None] = {}
    sucursal_cache: dict[tuple[int, str, str, int], Sucursal | None] = {}
    created = 0
    updated = 0
    skipped = 0
    terminated_early = False
    result_rows: list[dict] = []

    tx_cm = nullcontext() if dry_run else transaction.atomic()
    with tx_cm:
        for index, row in enumerate(rows, start=1):
            receta_id = row.get("receta_id")
            receta_name = str(row.get("receta") or "").strip()
            codigo_point = str(row.get("codigo_point") or "").strip()
            receta = _resolve_receta_bulk_ref(
                receta_id=receta_id,
                receta_name=receta_name,
                codigo_point=codigo_point,
                cache=receta_cache,
            )
            if receta is None:
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "receta_not_found",
                        "receta_id": int(receta_id or 0) or None,
                        "receta_input": receta_name or codigo_point,
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            sucursal_id = row.get("sucursal_id")
            sucursal_name = str(row.get("sucursal") or "").strip()
            sucursal_codigo = str(row.get("sucursal_codigo") or "").strip()
            has_sucursal_ref = bool(sucursal_id) or bool(sucursal_name) or bool(sucursal_codigo)
            sucursal = _resolve_sucursal_bulk_ref(
                sucursal_id=sucursal_id,
                sucursal_name=sucursal_name,
                sucursal_codigo=sucursal_codigo,
                default_sucursal=default_sucursal,
                cache=sucursal_cache,
            )
            if has_sucursal_ref and sucursal is None:
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "sucursal_not_found",
                        "receta_id": receta.id,
                        "receta": receta.nombre,
                        "sucursal_input": sucursal_codigo or sucursal_name or str(sucursal_id),
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            alcance_ui = str(row.get("alcance") or "mes").strip().lower()
            alcance_model = _ui_to_model_alcance(alcance_ui)
            periodo_default = _normalize_periodo_mes(row.get("periodo"))
            fecha_base_default = row.get("fecha_base") or timezone.localdate()
            periodo, fecha_inicio, fecha_fin = _resolve_solicitud_window(
                alcance=alcance_model,
                periodo_raw=row.get("periodo"),
                fecha_base_raw=row.get("fecha_base"),
                fecha_inicio_raw=row.get("fecha_inicio"),
                fecha_fin_raw=row.get("fecha_fin"),
                periodo_default=periodo_default,
                fecha_base_default=fecha_base_default,
            )
            cantidad = _to_decimal(row.get("cantidad"), default=Decimal("0"))
            if cantidad <= 0:
                skipped += 1
                result_rows.append(
                    {
                        "row": index,
                        "status": "ERROR",
                        "reason": "invalid_qty",
                        "receta_id": receta.id,
                        "receta": receta.nombre,
                        "cantidad": float(cantidad),
                    }
                )
                if stop_on_error:
                    terminated_early = True
                    break
                continue

            existing = SolicitudVenta.objects.filter(
                receta=receta,
                sucursal=sucursal,
                alcance=alcance_model,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
            ).first()

            previous_qty = _to_decimal(existing.cantidad, default=Decimal("0")) if existing else Decimal("0")
            if existing:
                new_qty = previous_qty + cantidad if modo == "accumulate" else cantidad
                action = "UPDATED"
                updated += 1
            else:
                new_qty = cantidad
                action = "CREATED"
                created += 1

            if not dry_run:
                if existing:
                    existing.periodo = periodo
                    existing.cantidad = new_qty
                    existing.fuente = fuente
                    existing.save(update_fields=["periodo", "cantidad", "fuente", "actualizado_en"])
                else:
                    SolicitudVenta.objects.create(
                        receta=receta,
                        sucursal=sucursal,
                        alcance=alcance_model,
                        periodo=periodo,
                        fecha_inicio=fecha_inicio,
                        fecha_fin=fecha_fin,
                        cantidad=new_qty,
                        fuente=fuente,
                    )

            result_rows.append(
                {
                    "row": index,
                    "status": action,
                    "receta_id": receta.id,
                    "receta": receta.nombre,
                    "sucursal_id": sucursal.id if sucursal else None,
                    "sucursal": sucursal.nombre if sucursal else "",
                    "alcance": alcance_model,
                    "periodo": periodo,
                    "fecha_inicio": str(fecha_inicio),
                    "fecha_fin": str(fecha_fin),
                    "cantidad_prev": float(previous_qty),
                    "cantidad_nueva": float(new_qty),
                }
            )

    error_count = sum(1 for row in result_rows if row.get("status") == "ERROR")
    return {
        "dry_run": dry_run,
        "mode": modo,
        "fuente": fuente,
        "terminated_early": terminated_early,
        "summary": {
            "total_rows": len(rows),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": error_count,
            "applied": 0 if dry_run else (created + updated),
        },
        "rows": result_rows[:top],
    }


class SolicitudVentaImportPreviewView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para previsualizar solicitudes de venta."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = SolicitudVentaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_solicitud_venta_bulk(ser.validated_data, dry_run_override=True)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        payload["preview"] = True
        return Response(payload, status=status.HTTP_200_OK)


class SolicitudVentaImportConfirmView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para confirmar importación de solicitudes de venta."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = SolicitudVentaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_solicitud_venta_bulk(ser.validated_data, dry_run_override=False)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        payload["preview"] = False
        return Response(payload, status=status.HTTP_200_OK)


class SolicitudVentaBulkUpsertView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para importar solicitudes de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = SolicitudVentaBulkSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            payload = _process_solicitud_venta_bulk(ser.validated_data)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(payload, status=status.HTTP_200_OK)


class ForecastEstadisticoView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.view_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para consultar pronóstico estadístico."},
                status=status.HTTP_403_FORBIDDEN,
            )
        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ser = ForecastEstadisticoRequestSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        periodo = _normalize_periodo_mes(data.get("periodo"))
        fecha_base = data.get("fecha_base") or timezone.localdate()
        alcance = data.get("alcance") or "mes"
        incluir_preparaciones = bool(data.get("incluir_preparaciones"))
        mix_adjustment_enabled = bool(data.get("mix_adjustment_enabled"))
        safety_pct = _to_decimal(data.get("safety_pct"), default=Decimal("0"))
        min_confianza_pct = _to_decimal(data.get("min_confianza_pct"), default=Decimal("0"))
        escenario_compare = str(data.get("escenario_compare") or "base").lower()
        top = int(data.get("top") or 120)

        sucursal = None
        sucursal_id = data.get("sucursal_id")
        if sucursal_id is not None:
            sucursal = Sucursal.objects.filter(pk=sucursal_id, activa=True).first()
            if sucursal is None:
                return Response(
                    {"detail": "Sucursal no encontrada o inactiva."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        result = _build_forecast_from_history(
            alcance=alcance,
            periodo=periodo,
            fecha_base=fecha_base,
            sucursal=sucursal,
            incluir_preparaciones=incluir_preparaciones,
            safety_pct=safety_pct,
            mix_adjustment_enabled=mix_adjustment_enabled,
        )
        result, filtered_conf = _filter_forecast_result_by_confianza(result, min_confianza_pct)
        result["min_confianza_pct"] = min_confianza_pct
        if not result.get("rows"):
            return Response(
                {"detail": "No hay forecast tras aplicar el filtro de confianza mínima."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        payload = _forecast_session_payload(result, top_rows=top)

        compare_payload = None
        if bool(data.get("include_solicitud_compare", True)):
            full_payload = _forecast_session_payload(result, top_rows=max(len(result.get("rows") or []), 1))
            compare_payload = _serialize_forecast_compare(
                _forecast_vs_solicitud_preview(full_payload, escenario=escenario_compare),
                top=top,
            )

        result_payload = {
            "scope": {
                "alcance": payload["alcance"],
                "periodo": payload["periodo"],
                "target_start": payload["target_start"],
                "target_end": payload["target_end"],
                "sucursal_nombre": payload["sucursal_nombre"],
                "sucursal_id": payload.get("sucursal_id"),
                "escenario_compare": escenario_compare,
                "mix_adjustment_enabled": mix_adjustment_enabled,
                "min_confianza_pct": _to_float(min_confianza_pct),
                "filtered_conf": filtered_conf,
            },
            "totals": payload["totals"],
            "rows": payload["rows"],
            "detail_rows": payload.get("detail_rows") or [],
            "detail_rows_total": int(payload.get("detail_rows_total") or 0),
            "model_meta": payload.get("model_meta") or {},
            "compare_solicitud": compare_payload,
        }
        if export_format:
            return _forecast_estadistico_export_response(result_payload, export_format)
        return Response(result_payload, status=status.HTTP_200_OK)


class ForecastEstadisticoGuardarView(APIView):
    permission_classes = [IsAuthenticated]

    _ESCENARIO_TO_KEY = {
        "base": "forecast_qty",
        "bajo": "forecast_low",
        "alto": "forecast_high",
    }

    def post(self, request):
        if not request.user.has_perm("recetas.change_pronosticoventa"):
            return Response(
                {"detail": "No tienes permisos para guardar pronóstico estadístico."},
                status=status.HTTP_403_FORBIDDEN,
            )
        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ser = ForecastEstadisticoGuardarSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        periodo = _normalize_periodo_mes(data.get("periodo"))
        fecha_base = data.get("fecha_base") or timezone.localdate()
        alcance = data.get("alcance") or "mes"
        incluir_preparaciones = bool(data.get("incluir_preparaciones"))
        mix_adjustment_enabled = bool(data.get("mix_adjustment_enabled"))
        safety_pct = _to_decimal(data.get("safety_pct"), default=Decimal("0"))
        min_confianza_pct = _to_decimal(data.get("min_confianza_pct"), default=Decimal("0"))
        top = int(data.get("top") or 120)
        escenario = str(data.get("escenario") or "base").lower()
        qty_key = self._ESCENARIO_TO_KEY.get(escenario, "forecast_qty")
        replace_existing = bool(data.get("replace_existing", True))
        fuente = (data.get("fuente") or "API_FORECAST_STAT").strip()[:40] or "API_FORECAST_STAT"

        sucursal = None
        sucursal_id = data.get("sucursal_id")
        if sucursal_id is not None:
            sucursal = Sucursal.objects.filter(pk=sucursal_id, activa=True).first()
            if sucursal is None:
                return Response(
                    {"detail": "Sucursal no encontrada o inactiva."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        result = _build_forecast_from_history(
            alcance=alcance,
            periodo=periodo,
            fecha_base=fecha_base,
            sucursal=sucursal,
            incluir_preparaciones=incluir_preparaciones,
            safety_pct=safety_pct,
            mix_adjustment_enabled=mix_adjustment_enabled,
        )
        result, filtered_conf = _filter_forecast_result_by_confianza(result, min_confianza_pct)
        result["min_confianza_pct"] = min_confianza_pct
        if not result.get("rows"):
            return Response(
                {"detail": "No hay historial suficiente para generar forecast en ese alcance/filtro."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        payload = _forecast_session_payload(result, top_rows=top)
        recetas_map = {r.id: r for r in Receta.objects.filter(id__in=[int(x["receta_id"]) for x in result["rows"]]).only("id", "nombre")}

        created = 0
        updated = 0
        skipped_existing = 0
        skipped_invalid = 0
        applied_rows: list[dict[str, Any]] = []

        with transaction.atomic():
            for row in result["rows"]:
                receta_id = int(row["receta_id"])
                receta = recetas_map.get(receta_id)
                if receta is None:
                    skipped_invalid += 1
                    continue

                qty = _to_decimal(row.get(qty_key))
                if qty <= 0:
                    skipped_invalid += 1
                    continue
                qty = qty.quantize(Decimal("0.001"))

                existing = PronosticoVenta.objects.filter(receta=receta, periodo=result["periodo"]).first()
                old_qty = _to_decimal(existing.cantidad if existing else 0).quantize(Decimal("0.001"))
                if existing is None:
                    PronosticoVenta.objects.create(
                        receta=receta,
                        periodo=result["periodo"],
                        cantidad=qty,
                        fuente=fuente,
                    )
                    created += 1
                    action = "create"
                else:
                    if not replace_existing:
                        skipped_existing += 1
                        continue
                    existing.cantidad = qty
                    existing.fuente = fuente
                    existing.save(update_fields=["cantidad", "fuente", "actualizado_en"])
                    updated += 1
                    action = "update"

                if len(applied_rows) < top:
                    applied_rows.append(
                        {
                            "receta_id": receta.id,
                            "receta": receta.nombre,
                            "escenario": escenario,
                            "cantidad_anterior": _to_float(old_qty),
                            "cantidad_nueva": _to_float(qty),
                            "accion": action,
                        }
                    )

        response_payload = {
            "scope": {
                "alcance": payload["alcance"],
                "periodo": payload["periodo"],
                "target_start": payload["target_start"],
                "target_end": payload["target_end"],
                "sucursal_nombre": payload["sucursal_nombre"],
                "sucursal_id": payload.get("sucursal_id"),
                "escenario": escenario,
                "qty_key": qty_key,
                "mix_adjustment_enabled": mix_adjustment_enabled,
                "min_confianza_pct": _to_float(min_confianza_pct),
                "filtered_conf": filtered_conf,
            },
            "totals": payload["totals"],
            "persisted": {
                "created": created,
                "updated": updated,
                "skipped_existing": skipped_existing,
                "skipped_invalid": skipped_invalid,
                "applied": created + updated,
            },
            "rows": payload["rows"],
            "detail_rows": payload.get("detail_rows") or [],
            "detail_rows_total": int(payload.get("detail_rows_total") or 0),
            "model_meta": payload.get("model_meta") or {},
            "applied_rows": applied_rows,
        }
        if export_format:
            return _forecast_estadistico_guardar_export_response(response_payload, export_format)
        return Response(response_payload, status=status.HTTP_200_OK)


class SolicitudVentaUpsertView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para capturar solicitudes de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )

        ser = SolicitudVentaUpsertSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        receta = get_object_or_404(Receta, pk=data["receta_id"])
        sucursal = None
        if data.get("sucursal_id") is not None:
            sucursal = Sucursal.objects.filter(pk=data["sucursal_id"], activa=True).first()
            if sucursal is None:
                return Response(
                    {"detail": "Sucursal no encontrada o inactiva."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        alcance = _ui_to_model_alcance(data.get("alcance"))
        periodo_default = _normalize_periodo_mes(data.get("periodo"))
        fecha_base_default = data.get("fecha_base") or timezone.localdate()
        periodo, fecha_inicio, fecha_fin = _resolve_solicitud_window(
            alcance=alcance,
            periodo_raw=data.get("periodo"),
            fecha_base_raw=data.get("fecha_base"),
            fecha_inicio_raw=data.get("fecha_inicio"),
            fecha_fin_raw=data.get("fecha_fin"),
            periodo_default=periodo_default,
            fecha_base_default=fecha_base_default,
        )
        cantidad = _to_decimal(data.get("cantidad"))
        fuente = (data.get("fuente") or "API_SOL_VENTAS").strip()[:40] or "API_SOL_VENTAS"

        with transaction.atomic():
            record, created = SolicitudVenta.objects.get_or_create(
                receta=receta,
                sucursal=sucursal,
                alcance=alcance,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                defaults={
                    "periodo": periodo,
                    "cantidad": cantidad,
                    "fuente": fuente,
                },
            )
            if not created:
                record.periodo = periodo
                record.cantidad = cantidad
                record.fuente = fuente
                record.save(update_fields=["periodo", "cantidad", "fuente", "actualizado_en"])

        forecast_ref = None
        validate_forecast = _parse_bool(request.GET.get("validate_forecast"), default=True)
        if validate_forecast:
            forecast_alcance = {
                SolicitudVenta.ALCANCE_MES: "mes",
                SolicitudVenta.ALCANCE_SEMANA: "semana",
                SolicitudVenta.ALCANCE_FIN_SEMANA: "fin_semana",
            }.get(alcance, "mes")
            forecast_incluir_preparaciones = _parse_bool(request.GET.get("incluir_preparaciones"), default=True)
            forecast_safety_pct = _to_decimal(request.GET.get("safety_pct"), default=Decimal("0"))
            forecast_min_confianza_pct = _to_decimal(request.GET.get("min_confianza_pct"), default=Decimal("0"))
            threshold_pct = _to_decimal(request.GET.get("umbral_desviacion_pct"), default=Decimal("20"))
            if threshold_pct < 0:
                threshold_pct = Decimal("20")

            try:
                forecast_result = _build_forecast_from_history(
                    alcance=forecast_alcance,
                    periodo=periodo,
                    fecha_base=fecha_inicio,
                    sucursal=sucursal,
                    incluir_preparaciones=forecast_incluir_preparaciones,
                    safety_pct=forecast_safety_pct,
                )
                forecast_result, _ = _filter_forecast_result_by_confianza(forecast_result, forecast_min_confianza_pct)
                row = next(
                    (r for r in (forecast_result.get("rows") or []) if int(r.get("receta_id") or 0) == receta.id),
                    None,
                )
                if row:
                    forecast_qty = _to_decimal(row.get("forecast_qty"))
                    delta_qty = (cantidad - forecast_qty).quantize(Decimal("0.001"))
                    variacion_pct = None
                    if forecast_qty > 0:
                        variacion_pct = ((delta_qty / forecast_qty) * Decimal("100")).quantize(Decimal("0.1"))

                    if forecast_qty <= 0 and cantidad > 0:
                        status_forecast = "SIN_BASE"
                    elif variacion_pct is None:
                        status_forecast = "OK"
                    elif variacion_pct > threshold_pct:
                        status_forecast = "SOBRE"
                    elif variacion_pct < (-threshold_pct):
                        status_forecast = "BAJO"
                    else:
                        status_forecast = "OK"

                    forecast_ref = {
                        "status": status_forecast,
                        "threshold_pct": _to_float(threshold_pct),
                        "forecast_qty": _to_float(forecast_qty),
                        "forecast_low": _to_float(row.get("forecast_low")),
                        "forecast_high": _to_float(row.get("forecast_high")),
                        "solicitud_qty": _to_float(cantidad),
                        "delta_qty": _to_float(delta_qty),
                        "variacion_pct": _to_float(variacion_pct) if variacion_pct is not None else None,
                        "confianza_pct": _to_float(row.get("confianza")),
                        "recomendacion": row.get("recomendacion") or "",
                    }
                else:
                    forecast_ref = {"status": "SIN_FORECAST"}
            except Exception:
                forecast_ref = {"status": "FORECAST_ERROR"}

        return Response(
            {
                "created": created,
                "id": record.id,
                "receta_id": record.receta_id,
                "receta": record.receta.nombre,
                "sucursal_id": record.sucursal_id,
                "sucursal": record.sucursal.nombre if record.sucursal_id else "",
                "alcance": record.alcance,
                "periodo": record.periodo,
                "fecha_inicio": str(record.fecha_inicio),
                "fecha_fin": str(record.fecha_fin),
                "cantidad": str(record.cantidad),
                "fuente": record.fuente,
                "forecast_ref": forecast_ref,
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )


class SolicitudVentaAplicarForecastView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.has_perm("recetas.change_planproduccion"):
            return Response(
                {"detail": "No tienes permisos para ajustar solicitudes de ventas."},
                status=status.HTTP_403_FORBIDDEN,
            )
        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format and export_format not in {"csv", "xlsx"}:
            return Response(
                {"detail": "export inválido. Usa csv o xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ser = SolicitudVentaAplicarForecastSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data

        sucursal = Sucursal.objects.filter(pk=data["sucursal_id"], activa=True).first()
        if sucursal is None:
            return Response(
                {"detail": "Sucursal no encontrada o inactiva."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        alcance = data.get("alcance") or "mes"
        periodo = _normalize_periodo_mes(data.get("periodo"))
        fecha_base = data.get("fecha_base") or timezone.localdate()
        incluir_preparaciones = bool(data.get("incluir_preparaciones"))
        safety_pct = _to_decimal(data.get("safety_pct"), default=Decimal("0"))
        min_confianza_pct = _to_decimal(data.get("min_confianza_pct"), default=Decimal("0"))
        escenario = str(data.get("escenario") or "base").lower()
        dry_run = bool(data.get("dry_run", False))
        max_variacion_pct = _to_decimal(data.get("max_variacion_pct"), default=Decimal("-1"))
        if max_variacion_pct < 0:
            max_variacion_pct = None
        top = int(data.get("top") or 120)

        result = _build_forecast_from_history(
            alcance=alcance,
            periodo=periodo,
            fecha_base=fecha_base,
            sucursal=sucursal,
            incluir_preparaciones=incluir_preparaciones,
            safety_pct=safety_pct,
        )
        result, _ = _filter_forecast_result_by_confianza(result, min_confianza_pct)
        if not result.get("rows"):
            return Response(
                {"detail": "No hay forecast tras aplicar el filtro de confianza mínima."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        full_payload = _forecast_session_payload(result, top_rows=max(len(result.get("rows") or []), 1))
        compare_raw = _forecast_vs_solicitud_preview(full_payload, escenario=escenario)
        if not compare_raw or not compare_raw.get("rows"):
            return Response(
                {"detail": "No hay filas de comparación forecast vs solicitud para aplicar ajuste."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        modo = data.get("modo") or "desviadas"
        receta_id = data.get("receta_id")
        rows = list(compare_raw["rows"])
        if modo == "sobre":
            rows = [r for r in rows if r.get("status") == "SOBRE"]
        elif modo == "bajo":
            rows = [r for r in rows if r.get("status") == "BAJO"]
        elif modo == "todas":
            rows = [r for r in rows if r.get("status") in {"SOBRE", "BAJO", "SIN_BASE", "OK"}]
        elif modo == "receta":
            rows = [r for r in rows if int(r.get("receta_id") or 0) == int(receta_id)]
        else:
            rows = [r for r in rows if r.get("status") in {"SOBRE", "BAJO"}]

        if not rows:
            return Response(
                {"detail": "No hay filas objetivo para el modo seleccionado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        model_alcance = _ui_to_model_alcance(alcance)
        target_start = result["target_start"]
        target_end = result["target_end"]
        result_periodo = result["periodo"]
        fuente = (data.get("fuente") or "API_FORECAST_ADJUST").strip()[:40] or "API_FORECAST_ADJUST"

        created = 0
        updated = 0
        skipped = 0
        skipped_cap = 0
        applied = 0
        adjusted_rows = []
        tx_cm = nullcontext() if dry_run else transaction.atomic()
        with tx_cm:
            for row in rows:
                receta = Receta.objects.filter(pk=row["receta_id"]).first()
                if receta is None:
                    skipped += 1
                    continue
                nueva_cantidad = _to_decimal(row.get("forecast_qty"))
                if nueva_cantidad < 0:
                    skipped += 1
                    continue
                record = SolicitudVenta.objects.filter(
                    receta=receta,
                    sucursal=sucursal,
                    alcance=model_alcance,
                    fecha_inicio=target_start,
                    fecha_fin=target_end,
                ).first()
                was_created = record is None
                old_qty = _to_decimal(record.cantidad if record is not None else 0)
                variacion_pct = None
                if old_qty > 0:
                    variacion_pct = ((nueva_cantidad - old_qty) / old_qty * Decimal("100")).quantize(Decimal("0.1"))
                if (max_variacion_pct is not None) and (variacion_pct is not None):
                    if abs(variacion_pct) > max_variacion_pct:
                        skipped += 1
                        skipped_cap += 1
                        continue
                if was_created:
                    if dry_run:
                        created += 1
                    else:
                        SolicitudVenta.objects.create(
                            receta=receta,
                            sucursal=sucursal,
                            alcance=model_alcance,
                            periodo=result_periodo,
                            fecha_inicio=target_start,
                            fecha_fin=target_end,
                            cantidad=nueva_cantidad,
                            fuente=fuente,
                        )
                        created += 1
                        applied += 1
                else:
                    if dry_run:
                        updated += 1
                    else:
                        record.periodo = result_periodo
                        record.cantidad = nueva_cantidad
                        record.fuente = fuente
                        record.save(update_fields=["periodo", "cantidad", "fuente", "actualizado_en"])
                        updated += 1
                        applied += 1

                adjusted_rows.append(
                    {
                        "receta_id": receta.id,
                        "receta": receta.nombre,
                        "anterior": _to_float(old_qty),
                        "nueva": _to_float(nueva_cantidad),
                        "variacion_pct": _to_float(variacion_pct) if variacion_pct is not None else None,
                        "accion": "create" if was_created else "update",
                        "status_before": row.get("status") or "",
                    }
                )

        compare_payload = _serialize_forecast_compare(compare_raw, top=top)
        response_payload = {
            "scope": {
                "alcance": alcance,
                "periodo": result_periodo,
                "target_start": str(target_start),
                "target_end": str(target_end),
                "sucursal_id": sucursal.id,
                "sucursal_nombre": f"{sucursal.codigo} - {sucursal.nombre}",
                "modo": modo,
                "escenario": escenario,
                "min_confianza_pct": _to_float(min_confianza_pct),
            },
            "updated": {
                "dry_run": dry_run,
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "skipped_cap": skipped_cap,
                "applied": applied,
            },
            "adjusted_rows": adjusted_rows[:top],
            "compare_solicitud": compare_payload,
        }
        if export_format:
            return _ventas_solicitud_aplicar_forecast_export_response(response_payload, export_format)
        return Response(response_payload, status=status.HTTP_200_OK)



