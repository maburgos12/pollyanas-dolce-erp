from __future__ import annotations

import argparse
import sqlite3
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_DB = ROOT / "tmp" / "db_recovered_20260324.sqlite3"
DEFAULT_OUTPUT_DIR = ROOT / "output" / "spreadsheet"

KNOWN_NON_COMMERCIAL_CATEGORIES_RAW = {
    "alegría",
    "granmark",
    "pillines",
    "coca-cola",
    "te",
    "accesorios de repostería",
    "industrias lec",
    "hielo y agua mar de cortéz",
    "regalos",
}

CLOSURE_EXCLUDED_NAME_TOKENS = (
    "vaso ",
    "vasos ",
    "letrero",
    "vela",
    "accesorio",
    "regalo",
    "topping ",
    " sin preparar",
)

CLOSURE_EXCLUDED_META_TOKENS = (
    "vaso preparado",
    "vasos preparados",
    "accesorio",
    "accesorios",
    "vela",
    "velas",
    "regalo",
    "regalos",
    "bebida",
    "bebidas",
    "letrero",
    "letreros",
)


@dataclass(frozen=True)
class RecipeInfo:
    receta_id: int
    codigo_point: str
    nombre: str
    tipo: str
    categoria: str
    familia: str


def normalize(value: Any) -> str:
    text = str(value or "").strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.lower().split())


def safe_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


KNOWN_NON_COMMERCIAL_CATEGORIES = {normalize(value) for value in KNOWN_NON_COMMERCIAL_CATEGORIES_RAW}


def month_bounds(period: str) -> tuple[date, date]:
    year, month = [int(part) for part in period.split("-")]
    first = date(year, month, 1)
    if month == 12:
        last = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last = date(year, month + 1, 1) - timedelta(days=1)
    return first, last


def fetch_rows(conn: sqlite3.Connection, sql: str, params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    cur = conn.execute(sql, params)
    columns = [col[0] for col in cur.description]
    return [dict(zip(columns, row)) for row in cur.fetchall()]


class PilotBuilder:
    def __init__(self, db_path: Path, period: str) -> None:
        self.db_path = db_path
        self.period = period
        self.period_start, self.period_end = month_bounds(period)
        self.period_start_str = self.period_start.isoformat()
        self.period_end_str = self.period_end.isoformat()
        self.conn = sqlite3.connect(str(db_path))
        self.conn.row_factory = sqlite3.Row

        self.branches: dict[int, dict[str, Any]] = {}
        self.products: dict[int, dict[str, Any]] = {}
        self.products_by_sku: dict[str, dict[str, Any]] = {}
        self.recipes_by_id: dict[int, RecipeInfo] = {}
        self.code_lookup: dict[str, RecipeInfo] = {}
        self.name_lookup: dict[str, RecipeInfo] = {}
        self.pending_match_lookup: dict[tuple[str, str], dict[str, Any]] = {}
        self.recipe_point_lookup: dict[tuple[str, str], bool | None] = {}

        self.raw_sales: list[dict[str, Any]] = []
        self.raw_production: list[dict[str, Any]] = []
        self.raw_transfer: list[dict[str, Any]] = []
        self.raw_waste: list[dict[str, Any]] = []
        self.raw_cedis_ledger: list[dict[str, Any]] = []
        self.excluded_rows: list[dict[str, Any]] = []
        self.pending_rows: list[dict[str, Any]] = []
        self.commercial_outside_cedis_rows: list[dict[str, Any]] = []

        self.comparativa: dict[str, dict[str, Any]] = {}
        self.period_summary: dict[str, int] = {}

    def load_references(self) -> None:
        for row in fetch_rows(
            self.conn,
            """
            SELECT id, name, normalized_name, external_id, erp_branch_id
            FROM pos_bridge_branches
            """,
        ):
            self.branches[row["id"]] = row

        for row in fetch_rows(
            self.conn,
            """
            SELECT id, sku, name, normalized_name, category
            FROM pos_bridge_products
            """,
        ):
            self.products[row["id"]] = row
            self.products_by_sku[normalize(row["sku"])] = row

        recipes = fetch_rows(
            self.conn,
            """
            SELECT id, codigo_point, nombre, nombre_normalizado, tipo, categoria, familia
            FROM recetas_receta
            """,
        )
        commercial_name_candidates: dict[str, list[RecipeInfo]] = defaultdict(list)
        all_code_candidates: dict[str, RecipeInfo] = {}

        for row in recipes:
            info = RecipeInfo(
                receta_id=row["id"],
                codigo_point=(row["codigo_point"] or "").strip(),
                nombre=row["nombre"],
                tipo=row["tipo"],
                categoria=row["categoria"],
                familia=row["familia"],
            )
            self.recipes_by_id[info.receta_id] = info
            if info.codigo_point:
                all_code_candidates[normalize(info.codigo_point)] = info
            if info.tipo == "PRODUCTO_FINAL":
                commercial_name_candidates[normalize(row["nombre_normalizado"] or row["nombre"])].append(info)
                if info.codigo_point:
                    self.code_lookup[normalize(info.codigo_point)] = info

        for row in fetch_rows(
            self.conn,
            """
            SELECT codigo_point, nombre_point, receta_id, activo
            FROM recetas_recetacodigopointalias
            WHERE activo = 1
            """,
        ):
            recipe = self.recipes_by_id.get(row["receta_id"])
            if recipe and recipe.tipo == "PRODUCTO_FINAL":
                self.code_lookup[normalize(row["codigo_point"])] = recipe
                commercial_name_candidates[normalize(row["nombre_point"])].append(recipe)

        for normalized_name, candidates in commercial_name_candidates.items():
            unique_ids = {candidate.receta_id for candidate in candidates}
            if len(unique_ids) == 1:
                self.name_lookup[normalized_name] = candidates[0]

        for row in fetch_rows(
            self.conn,
            """
            SELECT tipo, point_codigo, point_nombre, method, fuzzy_score, fuzzy_sugerencia, clasificacion_operativa
            FROM maestros_pointpendingmatch
            """,
        ):
            code_key = normalize(row["point_codigo"])
            name_key = normalize(row["point_nombre"])
            payload = {
                "tipo": row["tipo"],
                "method": row["method"],
                "fuzzy_score": row["fuzzy_score"],
                "fuzzy_sugerencia": row["fuzzy_sugerencia"],
                "clasificacion_operativa": row["clasificacion_operativa"],
            }
            if code_key:
                self.pending_match_lookup[("code", code_key)] = payload
            if name_key:
                self.pending_match_lookup[("name", name_key)] = payload

        recipe_presence = fetch_rows(
            self.conn,
            """
            SELECT point_code, normalized_name, MAX(CASE WHEN has_recipe_flag THEN 1 ELSE 0 END) AS has_recipe_point
            FROM pos_bridge_recipe_nodes
            GROUP BY point_code, normalized_name
            """,
        )
        for row in recipe_presence:
            code_key = normalize(row["point_code"])
            name_key = normalize(row["normalized_name"])
            flag = bool(row["has_recipe_point"]) if row["has_recipe_point"] is not None else None
            if code_key:
                self.recipe_point_lookup[("code", code_key)] = flag
            if name_key:
                self.recipe_point_lookup[("name", name_key)] = flag

    def resolve_recipe(self, receta_id: Any, code: str | None, name: str | None) -> tuple[RecipeInfo | None, str]:
        if receta_id:
            recipe = self.recipes_by_id.get(int(receta_id))
            if recipe:
                return recipe, "receta_id"
        code_key = normalize(code)
        if code_key and code_key in self.code_lookup:
            return self.code_lookup[code_key], "codigo_point"
        name_key = normalize(name)
        if name_key and name_key in self.name_lookup:
            return self.name_lookup[name_key], "nombre_normalizado"
        return None, ""

    def get_recipe_point_flag(self, code: str | None, name: str | None) -> str:
        code_key = normalize(code)
        name_key = normalize(name)
        code_match = self.recipe_point_lookup.get(("code", code_key))
        name_match = self.recipe_point_lookup.get(("name", name_key))
        if code_match is True or name_match is True:
            return "SI"
        if code_match is False or name_match is False:
            return "NO"
        return "DESCONOCIDO"

    def lookup_pending_hint(self, code: str | None, name: str | None) -> dict[str, Any] | None:
        code_key = normalize(code)
        name_key = normalize(name)
        return self.pending_match_lookup.get(("code", code_key)) or self.pending_match_lookup.get(("name", name_key))

    def lookup_noncommercial_category(self, code: str | None) -> str | None:
        code_key = normalize(code)
        product = self.products_by_sku.get(code_key)
        if not product:
            return None
        category = product.get("category") or ""
        return category if normalize(category) in KNOWN_NON_COMMERCIAL_CATEGORIES else None

    def is_recipe_executive_eligible(self, recipe: RecipeInfo) -> bool:
        if recipe.tipo != "PRODUCTO_FINAL":
            return False
        normalized_name = normalize(recipe.nombre)
        normalized_meta = " ".join(part for part in [normalize(recipe.categoria), normalize(recipe.familia)] if part)
        if normalized_name.startswith("sabor "):
            return False
        if normalized_name.endswith(" kg") or normalized_name.endswith(" kilo"):
            return False
        if any(token in normalized_name for token in CLOSURE_EXCLUDED_NAME_TOKENS):
            return False
        if any(token in normalized_meta for token in CLOSURE_EXCLUDED_META_TOKENS):
            return False
        return True

    def comparativa_bucket(self, recipe: RecipeInfo) -> dict[str, Any]:
        bucket = self.comparativa.get(recipe.codigo_point)
        if bucket is None:
            bucket = {
                "codigo_point": recipe.codigo_point,
                "producto": recipe.nombre,
                "clasificacion": "PRODUCTO_TERMINADO_COMERCIAL",
                "produccion_cedis_qty": 0.0,
                "transferencia_cedis_qty": 0.0,
                "venta_periodo": 0.0,
                "merma_periodo": 0.0,
                "entry_origin_branches": set(),
                "entry_destination_branches": set(),
                "warnings": set(),
            }
            self.comparativa[recipe.codigo_point] = bucket
        return bucket

    def register_pending(
        self,
        source_type: str,
        code: str | None,
        name: str | None,
        quantity: float,
        reason: str,
        branch: str = "",
    ) -> None:
        pending_hint = self.lookup_pending_hint(code, name)
        self.pending_rows.append(
            {
                "fuente": source_type,
                "codigo_point": code or "",
                "nombre_point": name or "",
                "tipo_detectado": pending_hint["clasificacion_operativa"] if pending_hint else "SIN_CLASIFICAR",
                "motivo_pendiente": reason,
                "tiene_receta_bom_point": self.get_recipe_point_flag(code, name),
                "siguiente_accion_sugerida": (
                    "Sincronizar receta/BOM de Point y homologar a receta ERP"
                    if self.get_recipe_point_flag(code, name) == "SI"
                    else "Clasificar manualmente y crear receta o alias en ERP"
                ),
                "branch_context": branch,
                "cantidad_detectada": quantity,
                "fuzzy_sugerencia": pending_hint["fuzzy_sugerencia"] if pending_hint else "",
                "fuzzy_score": pending_hint["fuzzy_score"] if pending_hint else "",
                "metodo_sugerencia": pending_hint["method"] if pending_hint else "",
            }
        )

    def register_excluded(
        self,
        source_type: str,
        code: str | None,
        name: str | None,
        quantity: float,
        reason: str,
        branch: str = "",
    ) -> None:
        self.excluded_rows.append(
            {
                "fuente": source_type,
                "codigo_point": code or "",
                "producto": name or "",
                "clasificacion": reason,
                "razon_exclusion": reason,
                "branch_context": branch,
                "cantidad": quantity,
            }
        )

    def classify_sales(self) -> None:
        rows = fetch_rows(
            self.conn,
            """
            SELECT s.sale_date, s.quantity, s.total_amount, s.branch_id, s.product_id, s.receta_id
            FROM pos_bridge_daily_sales s
            WHERE s.sale_date BETWEEN ? AND ?
            ORDER BY s.sale_date, s.branch_id, s.product_id
            """,
            (self.period_start_str, self.period_end_str),
        )
        self.period_summary["sales_rows"] = len(rows)

        for row in rows:
            product = self.products[row["product_id"]]
            branch = self.branches[row["branch_id"]]
            quantity = safe_float(row["quantity"])
            recipe, match_method = self.resolve_recipe(row["receta_id"], product["sku"], product["name"])
            category_norm = normalize(product["category"])

            raw = {
                "sale_date": row["sale_date"],
                "branch": branch["name"],
                "codigo_point": product["sku"],
                "producto_point": product["name"],
                "categoria_point": product["category"],
                "cantidad": quantity,
                "venta_total": safe_float(row["total_amount"]),
                "receta_id": row["receta_id"] or "",
                "canonical_code": "",
                "canonical_product": "",
                "pilot_status": "",
                "classification_reason": "",
                "match_method": match_method,
                "source_table": "pos_bridge_daily_sales",
            }

            if recipe and recipe.tipo == "PRODUCTO_FINAL" and self.is_recipe_executive_eligible(recipe):
                bucket = self.comparativa_bucket(recipe)
                bucket["venta_periodo"] += quantity
                raw.update(
                    {
                        "canonical_code": recipe.codigo_point,
                        "canonical_product": recipe.nombre,
                        "pilot_status": "incluido_comparativa",
                        "classification_reason": "venta_comercial_homologada",
                    }
                )
            elif recipe and recipe.tipo != "PRODUCTO_FINAL":
                reason = f"receta_no_comercial:{recipe.tipo}"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("VENTA", product["sku"], product["name"], quantity, reason, branch["name"])
            elif recipe and not self.is_recipe_executive_eligible(recipe):
                reason = "receta_fuera_scope_cierre_producto"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("VENTA", product["sku"], product["name"], quantity, reason, branch["name"])
            elif category_norm in KNOWN_NON_COMMERCIAL_CATEGORIES:
                reason = f"categoria_no_comercial:{product['category']}"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("VENTA", product["sku"], product["name"], quantity, reason, branch["name"])
            else:
                reason = "venta_sin_homologacion_comercial"
                raw.update({"pilot_status": "pendiente_homologacion", "classification_reason": reason})
                self.register_pending("VENTA", product["sku"], product["name"], quantity, reason, branch["name"])

            self.raw_sales.append(raw)

    def classify_production(self) -> None:
        rows = fetch_rows(
            self.conn,
            """
            SELECT production_date, branch_id, item_code, item_name, is_insumo, receta_id, produced_quantity
            FROM pos_bridge_production_lines
            WHERE production_date BETWEEN ? AND ?
            ORDER BY production_date, branch_id, item_code
            """,
            (self.period_start_str, self.period_end_str),
        )
        self.period_summary["production_rows"] = len(rows)

        for row in rows:
            branch = self.branches[row["branch_id"]]
            quantity = safe_float(row["produced_quantity"])
            recipe, match_method = self.resolve_recipe(row["receta_id"], row["item_code"], row["item_name"])
            raw = {
                "production_date": row["production_date"],
                "branch": branch["name"],
                "codigo_point": row["item_code"] or "",
                "producto_point": row["item_name"] or "",
                "is_insumo": int(bool(row["is_insumo"])),
                "cantidad": quantity,
                "receta_id": row["receta_id"] or "",
                "canonical_code": "",
                "canonical_product": "",
                "pilot_status": "",
                "classification_reason": "",
                "match_method": match_method,
                "source_table": "pos_bridge_production_lines",
            }

            if row["is_insumo"]:
                reason = "insumo"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("PRODUCCION", row["item_code"], row["item_name"], quantity, reason, branch["name"])
            elif recipe and recipe.tipo != "PRODUCTO_FINAL":
                reason = f"receta_no_comercial:{recipe.tipo}"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("PRODUCCION", row["item_code"], row["item_name"], quantity, reason, branch["name"])
            elif self.lookup_noncommercial_category(row["item_code"]):
                category = self.lookup_noncommercial_category(row["item_code"])
                reason = f"categoria_no_comercial:{category}"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("PRODUCCION", row["item_code"], row["item_name"], quantity, reason, branch["name"])
            elif recipe and recipe.tipo == "PRODUCTO_FINAL" and self.is_recipe_executive_eligible(recipe):
                raw.update({"canonical_code": recipe.codigo_point, "canonical_product": recipe.nombre})
                if branch["normalized_name"] == "cedis":
                    raw.update(
                        {
                            "pilot_status": "trazabilidad_entrada_cedis",
                            "classification_reason": "produccion_directa_cedis_raw",
                        }
                    )
                else:
                    reason = f"produccion_comercial_fuera_cedis:{branch['name']}"
                    raw.update({"pilot_status": "fuera_alcance_comparativa", "classification_reason": reason})
                    self.commercial_outside_cedis_rows.append(
                        {
                            "fuente": "PRODUCCION",
                            "codigo_point": recipe.codigo_point,
                            "producto": recipe.nombre,
                            "sucursal": branch["name"],
                            "cantidad": quantity,
                            "motivo": reason,
                        }
                    )
            elif recipe and not self.is_recipe_executive_eligible(recipe):
                reason = "receta_fuera_scope_cierre_producto"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("PRODUCCION", row["item_code"], row["item_name"], quantity, reason, branch["name"])
            else:
                reason = "produccion_sin_homologacion"
                raw.update({"pilot_status": "pendiente_homologacion", "classification_reason": reason})
                self.register_pending("PRODUCCION", row["item_code"], row["item_name"], quantity, reason, branch["name"])

            self.raw_production.append(raw)

    def classify_transfer(self) -> None:
        rows = fetch_rows(
            self.conn,
            """
            SELECT date(received_at) AS received_date, origin_branch_id, destination_branch_id, item_code, item_name,
                   is_insumo, receta_id, received_quantity, is_received
            FROM pos_bridge_transfer_lines
            WHERE is_received = 1
              AND date(received_at) BETWEEN ? AND ?
            ORDER BY received_date, origin_branch_id, destination_branch_id, item_code
            """,
            (self.period_start_str, self.period_end_str),
        )
        self.period_summary["transfer_rows"] = len(rows)

        for row in rows:
            origin_branch = self.branches[row["origin_branch_id"]]
            destination_branch = self.branches[row["destination_branch_id"]]
            quantity = safe_float(row["received_quantity"])
            recipe, match_method = self.resolve_recipe(row["receta_id"], row["item_code"], row["item_name"])
            raw = {
                "received_date": row["received_date"],
                "origin_branch": origin_branch["name"],
                "destination_branch": destination_branch["name"],
                "codigo_point": row["item_code"] or "",
                "producto_point": row["item_name"] or "",
                "is_insumo": int(bool(row["is_insumo"])),
                "cantidad": quantity,
                "receta_id": row["receta_id"] or "",
                "canonical_code": "",
                "canonical_product": "",
                "pilot_status": "",
                "classification_reason": "",
                "match_method": match_method,
                "source_table": "pos_bridge_transfer_lines",
            }

            if row["is_insumo"]:
                reason = "insumo"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("TRANSFERENCIA", row["item_code"], row["item_name"], quantity, reason, destination_branch["name"])
            elif recipe and recipe.tipo != "PRODUCTO_FINAL":
                reason = f"receta_no_comercial:{recipe.tipo}"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("TRANSFERENCIA", row["item_code"], row["item_name"], quantity, reason, destination_branch["name"])
            elif self.lookup_noncommercial_category(row["item_code"]):
                category = self.lookup_noncommercial_category(row["item_code"])
                reason = f"categoria_no_comercial:{category}"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("TRANSFERENCIA", row["item_code"], row["item_name"], quantity, reason, destination_branch["name"])
            elif recipe and recipe.tipo == "PRODUCTO_FINAL" and self.is_recipe_executive_eligible(recipe):
                raw.update({"canonical_code": recipe.codigo_point, "canonical_product": recipe.nombre})
                if destination_branch["normalized_name"] == "cedis":
                    raw.update(
                        {
                            "pilot_status": "trazabilidad_entrada_cedis",
                            "classification_reason": "transferencia_recibida_cedis_raw",
                        }
                    )
                else:
                    reason = f"transferencia_comercial_fuera_cedis:{destination_branch['name']}"
                    raw.update({"pilot_status": "fuera_alcance_comparativa", "classification_reason": reason})
                    self.commercial_outside_cedis_rows.append(
                        {
                            "fuente": "TRANSFERENCIA",
                            "codigo_point": recipe.codigo_point,
                            "producto": recipe.nombre,
                            "sucursal": destination_branch["name"],
                            "cantidad": quantity,
                            "motivo": reason,
                        }
                    )
            elif recipe and not self.is_recipe_executive_eligible(recipe):
                reason = "receta_fuera_scope_cierre_producto"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("TRANSFERENCIA", row["item_code"], row["item_name"], quantity, reason, destination_branch["name"])
            else:
                reason = "transferencia_sin_homologacion"
                raw.update({"pilot_status": "pendiente_homologacion", "classification_reason": reason})
                self.register_pending("TRANSFERENCIA", row["item_code"], row["item_name"], quantity, reason, destination_branch["name"])

            self.raw_transfer.append(raw)

    def classify_cedis_entries_from_ledger(self) -> None:
        rows = fetch_rows(
            self.conn,
            """
            SELECT date(m.fecha) AS movement_date, m.referencia, m.source_hash, m.cantidad,
                   r.id AS receta_id, r.codigo_point, r.nombre, r.tipo, r.categoria, r.familia
            FROM recetas_movimientoproductocedis m
            JOIN recetas_receta r ON r.id = m.receta_id
            WHERE date(m.fecha) BETWEEN ? AND ?
            ORDER BY date(m.fecha), m.id
            """,
            (self.period_start_str, self.period_end_str),
        )
        self.period_summary["cedis_ledger_rows"] = len(rows)

        for row in rows:
            quantity = safe_float(row["cantidad"])
            recipe = self.recipes_by_id.get(row["receta_id"])
            reference = row["referencia"] or ""
            if reference.startswith("POINT-PROD:"):
                source_kind = "produccion_cedis"
                source_label = "CEDIS"
            elif reference.startswith("POINT-TRANSFER:"):
                source_kind = "transferencia_cedis"
                source_label = "TRANSFERENCIA"
            else:
                source_kind = "otro"
                source_label = "OTRO"

            raw = {
                "movement_date": row["movement_date"],
                "referencia": reference,
                "source_hash": row["source_hash"] or "",
                "source_kind": source_kind,
                "codigo_point": row["codigo_point"] or "",
                "producto": row["nombre"] or "",
                "cantidad": quantity,
                "receta_tipo": row["tipo"] or "",
                "categoria": row["categoria"] or "",
                "familia": row["familia"] or "",
                "pilot_status": "",
                "classification_reason": "",
                "source_table": "recetas_movimientoproductocedis",
            }

            if recipe is None:
                reason = "ledger_sin_receta"
                raw.update({"pilot_status": "pendiente_homologacion", "classification_reason": reason})
                self.register_pending("LEDGER_CEDIS", row["codigo_point"], row["nombre"], quantity, reason, "CEDIS")
            elif recipe.tipo != "PRODUCTO_FINAL":
                reason = f"receta_no_comercial:{recipe.tipo}"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("LEDGER_CEDIS", row["codigo_point"], row["nombre"], quantity, reason, "CEDIS")
            elif self.lookup_noncommercial_category(row["codigo_point"]):
                category = self.lookup_noncommercial_category(row["codigo_point"])
                reason = f"categoria_no_comercial:{category}"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("LEDGER_CEDIS", row["codigo_point"], row["nombre"], quantity, reason, "CEDIS")
            elif not self.is_recipe_executive_eligible(recipe):
                reason = "receta_fuera_scope_cierre_producto"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("LEDGER_CEDIS", row["codigo_point"], row["nombre"], quantity, reason, "CEDIS")
            else:
                bucket = self.comparativa_bucket(recipe)
                if source_kind == "produccion_cedis":
                    bucket["produccion_cedis_qty"] += quantity
                elif source_kind == "transferencia_cedis":
                    bucket["transferencia_cedis_qty"] += quantity
                else:
                    bucket["warnings"].add("Movimiento ledger con referencia no clasificada")
                bucket["entry_origin_branches"].add(source_label)
                bucket["entry_destination_branches"].add("CEDIS")
                raw.update({"pilot_status": "incluido_comparativa", "classification_reason": source_kind})

            self.raw_cedis_ledger.append(raw)

    def classify_waste(self) -> None:
        rows = fetch_rows(
            self.conn,
            """
            SELECT date(movement_at) AS movement_date, branch_id, item_code, item_name, insumo_id, receta_id, quantity
            FROM pos_bridge_waste_lines
            WHERE date(movement_at) BETWEEN ? AND ?
            ORDER BY movement_date, branch_id, item_name
            """,
            (self.period_start_str, self.period_end_str),
        )
        self.period_summary["waste_rows"] = len(rows)

        for row in rows:
            branch = self.branches[row["branch_id"]]
            quantity = safe_float(row["quantity"])
            recipe, match_method = self.resolve_recipe(row["receta_id"], row["item_code"], row["item_name"])
            raw = {
                "movement_date": row["movement_date"],
                "branch": branch["name"],
                "codigo_point": row["item_code"] or "",
                "producto_point": row["item_name"] or "",
                "insumo_id": row["insumo_id"] or "",
                "cantidad": quantity,
                "receta_id": row["receta_id"] or "",
                "canonical_code": "",
                "canonical_product": "",
                "pilot_status": "",
                "classification_reason": "",
                "match_method": match_method,
                "source_table": "pos_bridge_waste_lines",
            }

            if row["insumo_id"]:
                reason = "insumo"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("MERMA", row["item_code"], row["item_name"], quantity, reason, branch["name"])
            elif recipe and recipe.tipo != "PRODUCTO_FINAL":
                reason = f"receta_no_comercial:{recipe.tipo}"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("MERMA", row["item_code"], row["item_name"], quantity, reason, branch["name"])
            elif self.lookup_noncommercial_category(row["item_code"]):
                category = self.lookup_noncommercial_category(row["item_code"])
                reason = f"categoria_no_comercial:{category}"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("MERMA", row["item_code"], row["item_name"], quantity, reason, branch["name"])
            elif recipe and recipe.tipo == "PRODUCTO_FINAL" and self.is_recipe_executive_eligible(recipe):
                bucket = self.comparativa_bucket(recipe)
                bucket["merma_periodo"] += quantity
                raw.update(
                    {
                        "canonical_code": recipe.codigo_point,
                        "canonical_product": recipe.nombre,
                        "pilot_status": "incluido_comparativa",
                        "classification_reason": "merma_comercial_homologada",
                    }
                )
            elif recipe and not self.is_recipe_executive_eligible(recipe):
                reason = "receta_fuera_scope_cierre_producto"
                raw.update({"pilot_status": "excluido_no_comercial", "classification_reason": reason})
                self.register_excluded("MERMA", row["item_code"], row["item_name"], quantity, reason, branch["name"])
            else:
                reason = "merma_sin_homologacion"
                raw.update({"pilot_status": "pendiente_homologacion", "classification_reason": reason})
                self.register_pending("MERMA", row["item_code"], row["item_name"], quantity, reason, branch["name"])

            self.raw_waste.append(raw)

    def build_comparativa_rows(self) -> list[dict[str, Any]]:
        rows = []
        for code in sorted(self.comparativa):
            bucket = self.comparativa[code]
            entrada = bucket["produccion_cedis_qty"] + bucket["transferencia_cedis_qty"]
            source_count = sum(1 for value in [bucket["produccion_cedis_qty"], bucket["transferencia_cedis_qty"]] if value > 0)
            if source_count == 2:
                bucket["warnings"].add("Alerta: producto con producción directa en CEDIS y transferencias a CEDIS; revisar posible doble conteo")
                fuente_entrada = "mixto"
            elif bucket["produccion_cedis_qty"] > 0:
                fuente_entrada = "produccion_cedis"
            elif bucket["transferencia_cedis_qty"] > 0:
                fuente_entrada = "transferencia_cedis"
            else:
                fuente_entrada = ""
            if entrada == 0 and bucket["venta_periodo"] > 0:
                bucket["warnings"].add("Venta en periodo sin entrada comercial registrada en CEDIS")
            rows.append(
                {
                    "codigo_point": bucket["codigo_point"],
                    "producto": bucket["producto"],
                    "clasificacion": bucket["clasificacion"],
                    "sucursal_origen_si_aplica": ", ".join(sorted(bucket["entry_origin_branches"])),
                    "sucursal_destino_si_aplica": ", ".join(sorted(bucket["entry_destination_branches"])),
                    "produccion_cedis": round(bucket["produccion_cedis_qty"], 4),
                    "transferencia_cedis": round(bucket["transferencia_cedis_qty"], 4),
                    "entrada_cedis": round(entrada, 4),
                    "venta_periodo": round(bucket["venta_periodo"], 4),
                    "merma_periodo": round(bucket["merma_periodo"], 4),
                    "saldo_teorico": round(entrada - bucket["venta_periodo"] - bucket["merma_periodo"], 4),
                    "fuente_entrada": fuente_entrada,
                    "observaciones": " | ".join(sorted(bucket["warnings"])),
                }
            )
        return rows

    def aggregate_rows(self, rows: list[dict[str, Any]], key_fields: list[str], sum_fields: list[str]) -> list[dict[str, Any]]:
        buckets: dict[tuple[Any, ...], dict[str, Any]] = {}
        for row in rows:
            key = tuple(row[field] for field in key_fields)
            if key not in buckets:
                buckets[key] = {field: row[field] for field in key_fields}
                for field in sum_fields:
                    buckets[key][field] = 0.0
                buckets[key]["ocurrencias"] = 0
            buckets[key]["ocurrencias"] += 1
            for field in sum_fields:
                buckets[key][field] += safe_float(row.get(field))
        return sorted(buckets.values(), key=lambda item: tuple(str(item[field]) for field in key_fields))

    def build_excluded_rows(self) -> list[dict[str, Any]]:
        return self.aggregate_rows(
            self.excluded_rows,
            ["fuente", "codigo_point", "producto", "clasificacion", "razon_exclusion", "branch_context"],
            ["cantidad"],
        )

    def build_pending_rows(self) -> list[dict[str, Any]]:
        return self.aggregate_rows(
            self.pending_rows,
            [
                "fuente",
                "codigo_point",
                "nombre_point",
                "tipo_detectado",
                "motivo_pendiente",
                "tiene_receta_bom_point",
                "siguiente_accion_sugerida",
                "branch_context",
                "fuzzy_sugerencia",
                "fuzzy_score",
                "metodo_sugerencia",
            ],
            ["cantidad_detectada"],
        )

    def build_traceability_rows(self) -> list[dict[str, Any]]:
        return [
            {
                "regla": "Periodo piloto",
                "fuente": "Snapshot SQLite recuperado",
                "criterio": f"Se usó {self.period} porque el PostgreSQL local no contiene las tablas ERP y los meses posteriores del snapshot muestran cobertura incompleta en producción/merma.",
                "nota_auditoria": "Supuesto explícito del piloto; no implica cierre ni aplicación al ERP.",
            },
            {
                "regla": "Ventas canónicas",
                "fuente": "pos_bridge_daily_sales",
                "criterio": "Solo entran ventas homologadas a receta PRODUCTO_FINAL por receta_id, alias de código Point o nombre normalizado único.",
                "nota_auditoria": "Ventas sin homologación comercial se mandan a pendientes; accesorios/bebidas se excluyen.",
            },
            {
                "regla": "Entrada comercial CEDIS por producción",
                "fuente": "recetas_movimientoproductocedis",
                "criterio": "La entrada principal se toma del ledger canónico de producto terminado en CEDIS. Las referencias POINT-PROD cuentan como producción central.",
                "nota_auditoria": "pos_bridge_production_lines queda solo como trazabilidad raw, no como base de suma principal.",
            },
            {
                "regla": "Entrada comercial CEDIS por transferencia",
                "fuente": "recetas_movimientoproductocedis",
                "criterio": "Las referencias POINT-TRANSFER dentro del ledger canónico cuentan como transferencias recibidas en CEDIS.",
                "nota_auditoria": "pos_bridge_transfer_lines queda como auditoría raw; la suma principal sale del ledger canónico.",
            },
            {
                "regla": "Mermas",
                "fuente": "pos_bridge_waste_lines",
                "criterio": "Solo entra merma homologada a receta PRODUCTO_FINAL; insumos y preparaciones se excluyen.",
                "nota_auditoria": "Mermas sin receta comercial quedan en pendientes y no se fuerzan.",
            },
            {
                "regla": "No duplicar CEDIS",
                "fuente": "Comparativa agregada",
                "criterio": "Si un SKU tiene producción CEDIS y transferencia a CEDIS en el mismo periodo, se marca alerta de revisión.",
                "nota_auditoria": "La alerta no afirma error; solo pide auditoría de doble conteo potencial.",
            },
            {
                "regla": "Homologación mínima",
                "fuente": "recetas_receta, recetas_recetacodigopointalias, pos_bridge_recipe_nodes, maestros_pointpendingmatch",
                "criterio": "Se prioriza receta_id; luego alias de código Point; luego nombre normalizado único.",
                "nota_auditoria": "Si no hay match determinístico, el movimiento no entra a comparativa y queda como pendiente.",
            },
            {
                "regla": "Exclusiones comerciales",
                "fuente": "pos_bridge_products.category y receta.tipo",
                "criterio": "Se excluyen categorías no comerciales observadas en ventas sin receta, y cualquier receta PREPARACION o movimiento marcado como insumo.",
                "nota_auditoria": "Esto evita mezclar accesorios, bebidas y preparados con producto terminado comercial.",
            },
        ]

    def build_resumen_rows(
        self,
        comparativa_rows: list[dict[str, Any]],
        excluded_rows: list[dict[str, Any]],
        pending_rows: list[dict[str, Any]],
    ) -> tuple[list[dict[str, Any]], list[str]]:
        total_entries = round(sum(row["entrada_cedis"] for row in comparativa_rows), 4)
        total_sales = round(sum(row["venta_periodo"] for row in comparativa_rows), 4)
        total_waste = round(sum(row["merma_periodo"] for row in comparativa_rows), 4)
        total_balance = round(sum(row["saldo_teorico"] for row in comparativa_rows), 4)
        mixed_sources = [row for row in comparativa_rows if row["fuente_entrada"] == "mixto"]
        no_entry_sales = [row for row in comparativa_rows if row["entrada_cedis"] == 0 and row["venta_periodo"] > 0]
        outside_cedis_qty = round(sum(row["cantidad"] for row in self.commercial_outside_cedis_rows), 4)

        hallazgos = [
            f"Periodo piloto seleccionado: {self.period} con fuente principal {self.db_path.name}.",
            f"Se detectaron {len(mixed_sources)} códigos con entrada mixta (producción CEDIS + transferencia a CEDIS).",
            f"Se dejaron fuera {len(self.commercial_outside_cedis_rows)} movimientos comerciales fuera de CEDIS por un total de {outside_cedis_qty} unidades.",
            f"Quedaron {len(pending_rows)} combinaciones pendientes de homologación y {len(excluded_rows)} combinaciones excluidas no comerciales.",
            f"Hay {len(no_entry_sales)} códigos con venta del periodo pero sin entrada comercial registrada en CEDIS en el mismo mes.",
        ]

        resumen = [
            {"metric": "periodo", "value": self.period},
            {"metric": "fecha_hora_corte", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
            {"metric": "fuente_datos", "value": str(self.db_path)},
            {"metric": "total_entradas_comerciales", "value": total_entries},
            {"metric": "total_ventas", "value": total_sales},
            {"metric": "total_mermas", "value": total_waste},
            {"metric": "saldo_teorico", "value": total_balance},
            {"metric": "total_codigos_comparados", "value": len(comparativa_rows)},
            {"metric": "total_codigos_excluidos", "value": len(excluded_rows)},
            {"metric": "total_pendientes_homologacion", "value": len(pending_rows)},
            {"metric": "rows_sales_fuente", "value": self.period_summary.get("sales_rows", 0)},
            {"metric": "rows_production_fuente", "value": self.period_summary.get("production_rows", 0)},
            {"metric": "rows_transfer_fuente", "value": self.period_summary.get("transfer_rows", 0)},
            {"metric": "rows_ledger_cedis_fuente", "value": self.period_summary.get("cedis_ledger_rows", 0)},
            {"metric": "rows_waste_fuente", "value": self.period_summary.get("waste_rows", 0)},
        ]
        return resumen, hallazgos

    def render_workbook(
        self,
        output_path: Path,
        resumen_rows: list[dict[str, Any]],
        hallazgos: list[str],
        comparativa_rows: list[dict[str, Any]],
        excluded_rows: list[dict[str, Any]],
        pending_rows: list[dict[str, Any]],
        trace_rows: list[dict[str, Any]],
    ) -> None:
        wb = Workbook()
        wb.remove(wb.active)

        summary_ws = wb.create_sheet("RESUMEN_EJECUTIVO")
        self.write_table(summary_ws, resumen_rows, title="RESUMEN_EJECUTIVO")
        start_row = len(resumen_rows) + 4
        summary_ws.cell(row=start_row, column=1, value="principales_hallazgos")
        summary_ws.cell(row=start_row, column=1).font = Font(bold=True)
        for index, finding in enumerate(hallazgos, start=1):
            summary_ws.cell(row=start_row + index, column=1, value=finding)

        self.write_table(wb.create_sheet("COMPARATIVA_PRODUCTO"), comparativa_rows, title="COMPARATIVA_PRODUCTO")
        self.write_table(wb.create_sheet("MOVIMIENTOS_PRODUCCION_RAW"), self.raw_production, title="MOVIMIENTOS_PRODUCCION_RAW")
        self.write_table(wb.create_sheet("MOVIMIENTOS_TRANSFER_RAW"), self.raw_transfer, title="MOVIMIENTOS_TRANSFER_RAW")
        self.write_table(wb.create_sheet("LEDGER_CEDIS_RAW"), self.raw_cedis_ledger, title="LEDGER_CEDIS_RAW")
        self.write_table(wb.create_sheet("MERMAS_RAW"), self.raw_waste, title="MERMAS_RAW")
        self.write_table(wb.create_sheet("VENTAS_RAW"), self.raw_sales, title="VENTAS_RAW")
        self.write_table(wb.create_sheet("EXCLUIDOS_NO_COMERCIALES"), excluded_rows, title="EXCLUIDOS_NO_COMERCIALES")
        self.write_table(wb.create_sheet("PENDIENTES_HOMOLOGACION"), pending_rows, title="PENDIENTES_HOMOLOGACION")
        self.write_table(wb.create_sheet("TRAZABILIDAD_REGLAS"), trace_rows, title="TRAZABILIDAD_REGLAS")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

    def write_table(self, ws, rows: list[dict[str, Any]], title: str = "") -> None:
        if title:
            ws.cell(row=1, column=1, value=title)
            ws.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="1F4E78")
        if not rows:
            ws.cell(row=3, column=1, value="Sin datos")
            return
        headers = list(rows[0].keys())
        header_row = 3
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center")
        for row_idx, row in enumerate(rows, start=header_row + 1):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(header))
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}"
        for index, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row in rows[:1500]:
                max_len = max(max_len, len(str(row.get(header, ""))))
            ws.column_dimensions[get_column_letter(index)].width = min(max(max_len + 2, 12), 48)

    def run(self, output_path: Path) -> dict[str, Any]:
        self.load_references()
        self.classify_sales()
        self.classify_production()
        self.classify_transfer()
        self.classify_cedis_entries_from_ledger()
        self.classify_waste()

        comparativa_rows = self.build_comparativa_rows()
        excluded_rows = self.build_excluded_rows()
        pending_rows = self.build_pending_rows()
        trace_rows = self.build_traceability_rows()
        resumen_rows, hallazgos = self.build_resumen_rows(comparativa_rows, excluded_rows, pending_rows)

        self.render_workbook(
            output_path=output_path,
            resumen_rows=resumen_rows,
            hallazgos=hallazgos,
            comparativa_rows=comparativa_rows,
            excluded_rows=excluded_rows,
            pending_rows=pending_rows,
            trace_rows=trace_rows,
        )

        total_entries = round(sum(row["entrada_cedis"] for row in comparativa_rows), 4)
        total_sales = round(sum(row["venta_periodo"] for row in comparativa_rows), 4)
        total_waste = round(sum(row["merma_periodo"] for row in comparativa_rows), 4)
        total_balance = round(sum(row["saldo_teorico"] for row in comparativa_rows), 4)

        return {
            "period": self.period,
            "output_path": str(output_path),
            "total_entries": total_entries,
            "total_sales": total_sales,
            "total_waste": total_waste,
            "total_balance": total_balance,
            "comparativa_codes": len(comparativa_rows),
            "excluded_codes": len(excluded_rows),
            "pending_codes": len(pending_rows),
            "mixed_source_codes": sum(1 for row in comparativa_rows if row["fuente_entrada"] == "mixto"),
            "commercial_outside_cedis_rows": len(self.commercial_outside_cedis_rows),
            "hallazgos": hallazgos,
        }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Genera un Excel piloto de comparativa producción/venta/merma.")
    parser.add_argument("--db", type=Path, default=DEFAULT_DB, help="Ruta al snapshot SQLite a analizar.")
    parser.add_argument("--period", default="2025-12", help="Periodo piloto en formato YYYY-MM.")
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Ruta de salida del archivo XLSX. Si no se pasa, se genera en output/spreadsheet/.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_path = args.output
    if output_path is None:
        DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = DEFAULT_OUTPUT_DIR / f"pilot_comparativa_producto_terminado_{args.period}.xlsx"

    builder = PilotBuilder(args.db, args.period)
    result = builder.run(output_path)

    print("PILOT_PERIOD=", result["period"])
    print("OUTPUT_PATH=", result["output_path"])
    print("TOTAL_ENTRIES=", result["total_entries"])
    print("TOTAL_SALES=", result["total_sales"])
    print("TOTAL_WASTE=", result["total_waste"])
    print("TOTAL_BALANCE=", result["total_balance"])
    print("COMPARATIVA_CODES=", result["comparativa_codes"])
    print("EXCLUDED_CODES=", result["excluded_codes"])
    print("PENDING_CODES=", result["pending_codes"])
    print("MIXED_SOURCE_CODES=", result["mixed_source_codes"])
    print("COMMERCIAL_OUTSIDE_CEDIS_ROWS=", result["commercial_outside_cedis_rows"])
    print("HALLAZGOS=")
    for item in result["hallazgos"]:
        print("-", item)


if __name__ == "__main__":
    main()
