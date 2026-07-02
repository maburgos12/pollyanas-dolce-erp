"""Tests de regresión del panel de Precio mínimo rentable (Monitor de Márgenes).

Semántica honesta de costo:
- FAB_COMPLETO cuando ProductoCostoOperativoMensual trae componentes reales
  (mano de obra + indirectos + empaque > 0).
- MP_FALLBACK (solo materia prima, rotulado) cuando no hay fabricación completa.
- REVENTA_HISTORICO desde ProductoReventaCostoHistoricoMensual.
- SIN_COSTO si no hay ninguna fuente.
Catálogo/precio/familia desde Point; margen meta según fuente; precio sugerido como piso.
"""

import json
from io import BytesIO
from datetime import date
from decimal import Decimal
from uuid import uuid4

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from core.models import Sucursal
from recetas.models import PoliticaMargenPrecio, Receta, RecetaAgrupacionAddon, RecetaCostoVersion
from pos_bridge.models import PointBranch, PointDailySale, PointProduct
from reportes.models import (
    ProductoCostoOperativoMensual,
    ProductoReventaCostoHistoricoMensual,
    ProductoSucursalContribucionMensual,
    RecetaCostoHistoricoMensual,
)
from control.models import MermaMensualSucursal


class PrecioSugeridoViewTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_ps", email="admin_ps@example.com", password="test12345",
        )
        self.client.force_login(self.user)
        self.sucursal = Sucursal.objects.create(
            codigo="S1", nombre="Sucursal Prueba", fecha_apertura=date(2020, 1, 1)
        )
        self.point_branch = PointBranch.objects.create(
            external_id="PB1", name="Sucursal Prueba", erp_branch=self.sucursal,
        )
        self.period = date.today().replace(day=1)

    # ---- helpers ----
    def _receta(self, nombre, codigo, modo=Receta.MODO_COSTEO_FABRICADO):
        return Receta.objects.create(
            nombre=nombre, codigo_point=codigo, tipo=Receta.TIPO_PRODUCTO_FINAL,
            modo_costeo=modo, familia="ERP", hash_contenido=f"h-{uuid4()}",
        )

    def _point(self, codigo, nombre, precio, categoria="Pastel Mediano", activo=True):
        return PointProduct.objects.create(
            external_id=f"ext-{codigo}-{uuid4().hex[:6]}", name=nombre, sku=codigo,
            category=categoria, precio=Decimal(str(precio)), precio_activo=activo,
        )

    def _operativo(self, receta, fab, mp=0, mo=0, ind=0, emp=0, unidades=10, periodo=None):
        return ProductoCostoOperativoMensual.objects.create(
            periodo=periodo or self.period, receta=receta,
            unidades_base=Decimal(str(unidades)), asp=Decimal("0"),
            costo_mp_unit=Decimal(str(mp)), mano_obra_prod_unit=Decimal(str(mo)),
            indirecto_prod_unit=Decimal(str(ind)), empaque_prod_unit=Decimal(str(emp)),
            costo_fabricacion_unit=Decimal(str(fab)),
        )

    def _mp_hist(self, receta, costo, periodo=None):
        return RecetaCostoHistoricoMensual.objects.create(
            periodo=periodo or self.period, receta=receta, costo_total=Decimal(str(costo)),
        )

    def _version(self, receta, costo):
        return RecetaCostoVersion.objects.create(
            receta=receta, version_num=1, hash_snapshot=f"s-{uuid4()}",
            costo_total=Decimal(str(costo)),
        )

    def _reventa_hist(self, point, costo, periodo=None):
        return ProductoReventaCostoHistoricoMensual.objects.create(
            periodo=periodo or self.period, producto_point=point,
            costo_promedio=Decimal(str(costo)),
        )

    def _sale_90d(self, point, receta, qty, net, sale_date=None):
        return PointDailySale.objects.create(
            branch=self.point_branch,
            product=point,
            receta=receta,
            sale_date=sale_date or date.today(),
            quantity=Decimal(str(qty)),
            total_amount=Decimal(str(net)),
            net_amount=Decimal(str(net)),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )

    def _fetch(self, **params):
        params.setdefault("meses", 6)
        resp = self.client.get(reverse("recetas:monitor_margenes_precio_sugerido"), params)
        self.assertEqual(resp.status_code, 200)
        return json.loads(resp.content)

    def _row(self, data, receta):
        return next((r for r in data["rows"] if r["receta_id"] == receta.id), None)

    def _row_by_name(self, data, name):
        return next((r for r in data["rows"] if r["nombre"] == name), None)

    # ---- fuente de costo ----
    def test_fab_completo_cuando_operativo_tiene_componentes(self):
        r = self._receta("Pastel Fab", "FAB1")
        self._point("FAB1", "Pastel Fab", precio=500)
        self._operativo(r, fab=100, mp=60, mo=20, ind=10, emp=10)
        row = self._row(self._fetch(), r)
        self.assertEqual(row["costo_fuente"], "FAB_COMPLETO")
        self.assertEqual(row["costo_completo"], "100.00")
        self.assertTrue(row["costo_incluye_mano_obra"])
        self.assertIsNotNone(row["breakdown"])
        self.assertEqual(row["breakdown"]["mo"], "20.000000")

    def test_mp_fallback_cuando_no_hay_operativo(self):
        r = self._receta("Pastel MP", "MP1")
        self._point("MP1", "Pastel MP", precio=500)
        self._mp_hist(r, 80)
        row = self._row(self._fetch(), r)
        self.assertEqual(row["costo_fuente"], "MP_FALLBACK")
        self.assertEqual(row["costo_completo"], "80.00")
        self.assertFalse(row["costo_incluye_mano_obra"])
        self.assertIsNone(row["breakdown"])

    def test_operativo_sin_componentes_es_mp_fallback(self):
        # costo_fab>0 pero mo/ind/emp=0 -> NO es fabricación completa
        r = self._receta("Pastel MPop", "MP2")
        self._point("MP2", "Pastel MPop", precio=500)
        self._operativo(r, fab=90, mp=90, mo=0, ind=0, emp=0)
        row = self._row(self._fetch(), r)
        self.assertEqual(row["costo_fuente"], "MP_FALLBACK")
        self.assertEqual(row["costo_completo"], "90.00")

    def test_version_es_ultimo_fallback(self):
        r = self._receta("Pastel Ver", "VER1")
        self._point("VER1", "Pastel Ver", precio=500)
        self._version(r, 70)
        row = self._row(self._fetch(), r)
        self.assertEqual(row["costo_fuente"], "MP_FALLBACK")
        self.assertEqual(row["costo_completo"], "70.00")

    def test_reventa_usa_historico(self):
        r = self._receta("Agua", "RV1", modo=Receta.MODO_COSTEO_REVENTA)
        p = self._point("RV1", "Agua", precio=30, categoria="Granmark")
        self._reventa_hist(p, 12)
        row = self._row(self._fetch(), r)
        self.assertEqual(row["costo_fuente"], "REVENTA_HISTORICO")
        self.assertEqual(row["costo_completo"], "12.00")
        self.assertTrue(row["es_reventa"])
        self.assertEqual(row["margen_meta"], "30")

    def test_sin_costo_cuando_no_hay_fuente(self):
        r = self._receta("Sin Costo", "SC1")
        self._point("SC1", "Sin Costo", precio=200)
        row = self._row(self._fetch(), r)
        self.assertEqual(row["estado"], "SIN_COSTO")
        self.assertEqual(row["costo_fuente"], "SIN_COSTO")

    # ---- estados / margen 55% ----
    def test_margen_mayor_55_es_ok_y_no_baja(self):
        r = self._receta("Sano", "OK1")
        self._point("OK1", "Sano", precio=500)
        self._operativo(r, fab=95, mp=55, mo=20, ind=10, emp=10)
        row = self._row(self._fetch(), r)
        # margen (500-95)/500 = 81% >= 55 -> OK; sugerido 95/0.45=211->215 < 500
        self.assertEqual(row["estado"], "OK")
        self.assertEqual(row["precio_sugerido"], "215.00")
        self.assertEqual(row["precio_minimo_rentable"], "215.00")
        self.assertEqual(row["precio_recomendado"], "500.00")
        self.assertIsNone(row["falta_subir_pct"])

    def test_margen_menor_55_requiere_ajuste(self):
        r = self._receta("Ajuste", "AJ1")
        self._point("AJ1", "Ajuste", precio=150)
        self._operativo(r, fab=100, mp=60, mo=20, ind=10, emp=10)
        row = self._row(self._fetch(), r)
        # margen (150-100)/150 = 33% < 55 -> AJUSTE; sugerido 100/0.45=222->225
        self.assertEqual(row["estado"], "AJUSTE")
        self.assertEqual(row["precio_sugerido"], "225.00")
        self.assertIsNotNone(row["falta_subir_pct"])

    def test_precio_menor_costo_es_critico(self):
        r = self._receta("Critico", "CR1")
        self._point("CR1", "Critico", precio=90)
        self._operativo(r, fab=100, mp=60, mo=20, ind=10, emp=10)
        row = self._row(self._fetch(), r)
        self.assertEqual(row["estado"], "CRITICO")
        self.assertEqual(row["margen_actual"], "-11.1")

    # ---- catálogo Point ----
    def test_inactivo_se_excluye(self):
        r = self._receta("Inactivo", "IN1")
        self._point("IN1", "Inactivo", precio=200, activo=False)
        self._operativo(r, fab=50, mp=30, mo=10, ind=5, emp=5)
        self.assertIsNone(self._row(self._fetch(), r))

    def test_archivado_con_precio_vivo_se_excluye(self):
        # Producto desactivado en Point (active=False) que aún conserva precio
        # vigente NO debe aparecer en el catálogo de "productos activos de Point".
        r = self._receta("Archivado", "ARCH1")
        p = self._point("ARCH1", "Archivado", precio=200)
        p.active = False
        p.save(update_fields=["active"])
        self._operativo(r, fab=50, mp=30, mo=10, ind=5, emp=5)
        self.assertIsNone(self._row(self._fetch(), r))

    def test_familia_y_precio_vienen_de_point(self):
        r = self._receta("Conf", "CF1")
        self._point("CF1", "Conf", precio=400, categoria="Pay Grande")
        self._operativo(r, fab=100, mp=60, mo=20, ind=10, emp=10)
        row = self._row(self._fetch(), r)
        self.assertEqual(row["familia_point"], "Pay Grande")
        self.assertEqual(row["precio_actual"], "400.00")

    def test_activo_sin_precio_point_no_desaparece(self):
        r = self._receta("Mini activo sin precio", "MINI1")
        p = self._point("MINI1", "Mini activo sin precio", precio=0, categoria="Pastel Mini")
        p.precio = None
        p.save(update_fields=["precio"])
        self._operativo(r, fab=50, mp=50, mo=0, ind=0, emp=0)
        row = self._row(self._fetch(), r)
        self.assertIsNotNone(row)
        self.assertEqual(row["familia_point"], "Pastel Mini")
        self.assertEqual(row["estado"], "SIN_PRECIO")
        self.assertIsNone(row["precio_actual"])

    def test_filtro_familia_point(self):
        r1 = self._receta("A", "A1"); self._point("A1", "A", 300, categoria="Pastel Mediano")
        self._operativo(r1, fab=100, mp=60, mo=20, ind=10, emp=10)
        r2 = self._receta("B", "B1"); self._point("B1", "B", 300, categoria="Pay Grande")
        self._operativo(r2, fab=100, mp=60, mo=20, ind=10, emp=10)
        data = self._fetch(familia="Pay Grande")
        self.assertIsNone(self._row(data, r1))
        self.assertIsNotNone(self._row(data, r2))

    # ---- addons ----
    def test_addon_se_lista_como_combinacion_base_mas_sabor(self):
        base = self._receta("Pay de Queso Grande", "0001")
        self._point("0001", "Pay de Queso Grande", precio=500)
        self._operativo(base, fab=178, mp=178, mo=0, ind=0, emp=0, unidades=800)
        self._mp_hist(base, 178)  # MP fallback de la base
        addon = self._receta("Sabor Fresa Grande", "SFRESAG")
        self._operativo(addon, fab=73, mp=73, mo=0, ind=0, emp=0, unidades=800)
        RecetaAgrupacionAddon.objects.create(
            base_receta=base, addon_receta=addon, addon_codigo_point="SFRESAG",
            addon_nombre_point="Sabor Fresa Grande", status=RecetaAgrupacionAddon.STATUS_APPROVED,
        )
        data = self._fetch()
        self.assertIsNone(self._row(data, addon))  # addon excluido
        self.assertIsNone(self._row_by_name(data, "Pay de Queso Grande"))  # base excluida si hay producto completo

        combo = self._row_by_name(data, "Pay de Queso Grande + Sabor Fresa Grande")
        self.assertIsNotNone(combo)
        self.assertEqual(combo["codigo_point"], "0001 + SFRESAG")
        self.assertEqual(combo["costo_completo"], "251.00")
        self.assertEqual(combo["addon_cost"], "73.00")
        self.assertEqual(combo["precio_actual"], "500.00")
        self.assertTrue(combo["tiene_sabores"])
        self.assertEqual(combo["precio_sugerido"], "720.00")

    # ---- export ----
    def test_export_csv(self):
        r = self._receta("CSV", "CS1")
        self._point("CS1", "CSV", precio=300)
        self._operativo(r, fab=100, mp=60, mo=20, ind=10, emp=10)
        resp = self.client.get(
            reverse("recetas:monitor_margenes_precio_sugerido"), {"meses": 6, "export": "csv"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertIn("text/csv", resp["Content-Type"])

    def test_export_xlsx_con_branding_y_agrupado_por_familia(self):
        r1 = self._receta("Pastel A", "XA1")
        self._point("XA1", "Pastel A", precio=300, categoria="Pastel Mediano")
        self._operativo(r1, fab=100, mp=60, mo=20, ind=10, emp=10)
        r2 = self._receta("Pay B", "XB1")
        self._point("XB1", "Pay B", precio=300, categoria="Pay Grande")
        self._operativo(r2, fab=100, mp=60, mo=20, ind=10, emp=10)

        resp = self.client.get(
            reverse("recetas:monitor_margenes_precio_sugerido"), {"meses": 6, "export": "xlsx"},
        )

        self.assertEqual(resp.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resp["Content-Type"],
        )
        wb = load_workbook(BytesIO(resp.content))
        ws = wb["Precio sugerido"]
        self.assertEqual(ws["A5"].value, "Pollyana's Dolce - Precio minimo y recomendado")
        family_headers = [
            cell.value for row in ws.iter_rows(min_row=13, max_col=1) for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("Familia:")
        ]
        self.assertIn("Familia: Pastel Mediano (1 productos)", family_headers)
        self.assertIn("Familia: Pay Grande (1 productos)", family_headers)
        self.assertTrue(any(drawing.anchor._from.row == 0 for drawing in ws._images))

    # ---- margen meta por fuente ----
    def test_margen_meta_55_para_fab_completo(self):
        r = self._receta("Fab", "MF1")
        self._point("MF1", "Fab", precio=500)
        self._operativo(r, fab=100, mp=60, mo=20, ind=10, emp=10)
        data = self._fetch()
        self.assertEqual(data["margen_meta_fab"], "55")
        self.assertEqual(data["margen_meta_mp"], "65")
        self.assertEqual(self._row(data, r)["margen_meta"], "55")

    def test_mp_fallback_usa_margen_meta_65(self):
        # Solo MP: meta 65%. Precio 200, costo 80 -> margen 60% (>=55 pero <65) -> AJUSTE.
        r = self._receta("SoloMP", "MM1")
        self._point("MM1", "SoloMP", precio=200)
        self._mp_hist(r, 80)
        row = self._row(self._fetch(), r)
        self.assertEqual(row["margen_meta"], "65")
        self.assertEqual(row["estado"], "AJUSTE")
        # sugerido = 80 / (1 - 0.65) = 228.57 -> redondeo techo $5 = 230
        self.assertEqual(row["precio_sugerido"], "230.00")

    def test_reventa_usa_margen_meta_30_y_no_fuerza_markup_fabricado(self):
        r = self._receta("Caja carton", "CAJA1", modo=Receta.MODO_COSTEO_REVENTA)
        p = self._point("CAJA1", "Caja carton", precio=60, categoria="Accesorios")
        self._reventa_hist(p, Decimal("19.87"))
        data = self._fetch()
        row = self._row(data, r)
        self.assertEqual(data["margen_meta_reventa"], "30")
        self.assertEqual(row["margen_meta"], "30")
        self.assertEqual(row["estado"], "OK")
        # sugerido = 19.87 / (1 - 0.30) = 28.39 -> redondeo techo $5 = 30
        self.assertEqual(row["precio_sugerido"], "30.00")
        self.assertIsNone(row["falta_subir_pct"])

    def test_politica_familia_fuente_sobrescribe_meta_mp(self):
        PoliticaMargenPrecio.objects.create(
            fuente_costo=PoliticaMargenPrecio.FUENTE_MP_FALLBACK,
            familia_point="Pastel Mediano",
            margen_meta_pct=Decimal("55.00"),
            subida_maxima_pct=Decimal("25.00"),
            prioridad=1,
        )
        r = self._receta("SoloMP Familiar", "MMF1")
        self._point("MMF1", "SoloMP Familiar", precio=200, categoria="Pastel Mediano")
        self._mp_hist(r, 80)

        row = self._row(self._fetch(), r)

        self.assertEqual(row["margen_meta"], "55")
        self.assertEqual(row["estado"], "OK")
        self.assertEqual(row["accion_sugerida"], "MANTENER")
        self.assertEqual(row["precio_sugerido"], "180.00")

    def test_subida_mayor_a_politica_propone_recomendado_gradual(self):
        PoliticaMargenPrecio.objects.create(
            fuente_costo=PoliticaMargenPrecio.FUENTE_FAB_COMPLETO,
            familia_point="Pastel Mediano",
            margen_meta_pct=Decimal("55.00"),
            subida_maxima_pct=Decimal("10.00"),
            prioridad=1,
        )
        r = self._receta("Fab Riesgo", "RIESGO1")
        self._point("RIESGO1", "Fab Riesgo", precio=150, categoria="Pastel Mediano")
        self._operativo(r, fab=100, mp=60, mo=20, ind=10, emp=10)

        row = self._row(self._fetch(), r)

        self.assertEqual(row["estado"], "AJUSTE")
        self.assertEqual(row["precio_sugerido"], "225.00")
        self.assertEqual(row["precio_minimo_rentable"], "225.00")
        self.assertEqual(row["precio_recomendado"], "165.00")
        self.assertEqual(row["subida_recomendada_pct"], "10.0")
        self.assertEqual(row["brecha_precio_minimo"], "60.00")
        self.assertEqual(row["accion_sugerida"], "SUBIR_GRADUAL")

    def test_ventas_90d_e_impacto_recomendado_usan_ventas_point(self):
        PoliticaMargenPrecio.objects.create(
            fuente_costo=PoliticaMargenPrecio.FUENTE_FAB_COMPLETO,
            familia_point="Pastel Mediano",
            margen_meta_pct=Decimal("55.00"),
            subida_maxima_pct=Decimal("10.00"),
            prioridad=1,
        )
        r = self._receta("Fab Volumen", "VOL1")
        point = self._point("VOL1", "Fab Volumen", precio=150, categoria="Pastel Mediano")
        self._operativo(r, fab=100, mp=60, mo=20, ind=10, emp=10)
        self._sale_90d(point, r, qty=12, net=Decimal("1800.00"))

        row = self._row(self._fetch(), r)

        self.assertEqual(row["ventas_90d_qty"], "12.000")
        self.assertEqual(row["ventas_90d_neto"], "1800.00")
        self.assertEqual(row["ventas_90d_asp"], "150.00")
        self.assertEqual(row["impacto_venta_90d"], "180.00")

    def test_monitor_muestra_inputs_para_editar_metas(self):
        resp = self.client.get(reverse("recetas:monitor_margenes"))

        self.assertEqual(resp.status_code, 200)
        html = resp.content.decode()
        self.assertIn('data-source="fab"', html)
        self.assertIn('data-source="mp"', html)
        self.assertIn('data-source="reventa"', html)
        self.assertIn("Guardar metas", html)

    def test_monitor_muestra_familias_point_activas_aunque_no_tengan_precio(self):
        p = self._point("PMINI1", "Pastel Mini Demo", precio=0, categoria="Pastel Mini")
        p.precio = None
        p.save(update_fields=["precio"])

        resp = self.client.get(reverse("recetas:monitor_margenes"))

        self.assertEqual(resp.status_code, 200)
        self.assertIn("Pastel Mini", resp.content.decode())

    def test_guardar_politicas_base_desde_panel_actualiza_calculo(self):
        r = self._receta("SoloMP Panel", "MMPANEL1")
        self._point("MMPANEL1", "SoloMP Panel", precio=200)
        self._mp_hist(r, 80)

        resp = self.client.post(
            reverse("recetas:monitor_margenes_politicas_precio"),
            data=json.dumps({
                "fab": {"margen_meta_pct": "50", "subida_maxima_pct": "30"},
                "mp": {"margen_meta_pct": "55", "subida_maxima_pct": "25"},
                "reventa": {"margen_meta_pct": "30", "subida_maxima_pct": "15"},
            }),
            content_type="application/json",
        )

        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.content)
        self.assertTrue(data["ok"])
        self.assertEqual(data["politicas_base"]["mp"]["margen_meta_pct"], "55")
        row = self._row(self._fetch(), r)
        self.assertEqual(row["margen_meta"], "55")


class RentabilidadProductoTests(PrecioSugeridoViewTests):
    """Campos de rentabilidad agregados al JSON de Precio sugerido para la
    pestaña Rentabilidad: contribución real, merma real, gasto comercial."""

    def _contribucion(self, receta, *, contribucion_total, gasto_comercial_total, unidades, periodo=None):
        return ProductoSucursalContribucionMensual.objects.create(
            periodo=periodo or self.period,
            receta=receta,
            sucursal=self.sucursal,
            unidades_vendidas=Decimal(str(unidades)),
            gasto_comercial_total=Decimal(str(gasto_comercial_total)),
            gasto_comercial_unit=(Decimal(str(gasto_comercial_total)) / Decimal(str(unidades))) if unidades else Decimal("0"),
            contribucion_total=Decimal(str(contribucion_total)),
        )

    def _merma(self, receta, *, costo, periodo=None):
        return MermaMensualSucursal.objects.create(
            periodo=periodo or self.period,
            sucursal=self.sucursal,
            receta=receta,
            nombre_producto=receta.nombre,
            costo_merma=Decimal(str(costo)),
        )

    def test_utilidad_estimada_resta_merma_real(self):
        r = self._receta("Pastel Rentable", "RENT1")
        self._point("RENT1", "Pastel Rentable", precio=500)
        self._operativo(r, fab=100, mp=60, mo=20, ind=10, emp=10)
        self._contribucion(r, contribucion_total=Decimal("4000.00"), gasto_comercial_total=Decimal("500.00"), unidades=100)
        self._merma(r, costo=Decimal("100.00"))

        row = self._row(self._fetch(), r)

        self.assertEqual(row["contribucion_total"], "4000.00")
        self.assertEqual(row["gasto_comercial_unit"], "5.00")
        self.assertTrue(row["gasto_comercial_desglose_disponible"])
        self.assertEqual(row["merma_unit"], "1.00")
        # utilidad_estimada_unit = (contribucion_total - merma_total) / unidades = (4000-100)/100
        self.assertEqual(row["utilidad_estimada_unit"], "39.00")

    def test_sin_filas_de_merma_es_cero_no_faltante(self):
        r = self._receta("Pastel Sin Merma", "RENT2")
        self._point("RENT2", "Pastel Sin Merma", precio=300)
        self._operativo(r, fab=80, mp=50, mo=15, ind=10, emp=5)
        self._contribucion(r, contribucion_total=Decimal("1000.00"), gasto_comercial_total=Decimal("200.00"), unidades=50)

        row = self._row(self._fetch(), r)

        self.assertEqual(row["merma_unit"], "0.00")
        self.assertIsNotNone(row["utilidad_estimada_unit"])

    def test_sin_contribucion_declara_gasto_comercial_no_disponible(self):
        r = self._receta("Pastel Sin Contribucion", "RENT3")
        self._point("RENT3", "Pastel Sin Contribucion", precio=300)
        self._mp_hist(r, 80)

        row = self._row(self._fetch(), r)

        self.assertEqual(row["contribucion_total"], "0.00")
        self.assertIsNone(row["gasto_comercial_unit"])
        self.assertFalse(row["gasto_comercial_desglose_disponible"])
        self.assertIsNone(row["utilidad_estimada_unit"])

    def test_sin_costo_no_rompe_calculo_rentabilidad(self):
        r = self._receta("Pastel Sin Costo Alguno", "RENT4")
        self._point("RENT4", "Pastel Sin Costo Alguno", precio=100)

        row = self._row(self._fetch(), r)

        self.assertEqual(row["costo_fuente"], "SIN_COSTO")
        self.assertEqual(row["merma_unit"], "0.00")

    def test_suma_contribucion_entre_sucursales_de_la_ventana(self):
        r = self._receta("Pastel Multisucursal", "RENT5")
        self._point("RENT5", "Pastel Multisucursal", precio=400)
        self._operativo(r, fab=90, mp=60, mo=15, ind=10, emp=5)
        otra_sucursal = Sucursal.objects.create(codigo="S2", nombre="Otra Sucursal", fecha_apertura=date(2020, 1, 1))
        self._contribucion(r, contribucion_total=Decimal("1000.00"), gasto_comercial_total=Decimal("100.00"), unidades=20)
        ProductoSucursalContribucionMensual.objects.create(
            periodo=self.period, receta=r, sucursal=otra_sucursal,
            unidades_vendidas=Decimal("30"), gasto_comercial_total=Decimal("150.00"),
            gasto_comercial_unit=Decimal("5.00"), contribucion_total=Decimal("1500.00"),
        )

        row = self._row(self._fetch(), r)

        self.assertEqual(row["contribucion_total"], "2500.00")
        self.assertEqual(row["estado"], "OK")
