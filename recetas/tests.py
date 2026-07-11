from datetime import date, timedelta
from calendar import monthrange
from decimal import Decimal
import os
import tempfile
from io import BytesIO, StringIO
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group, Permission
from django.core.cache import cache
from django.core.management import call_command
from django.core.management.base import CommandError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import OperationalError
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from core.access import ROLE_COMPRAS, ROLE_VENTAS
from core.models import UserProfile
from compras.models import OrdenCompra, RecepcionCompra, SolicitudCompra
from core.models import Sucursal
from maestros.models import CostoInsumo, Insumo, Proveedor, UnidadMedida
from recetas.management.commands.importar_ventas_point_archivos import Command as ImportarVentasPointArchivosCommand
from recetas.models import (
    CostoDriver,
    LineaReceta,
    PlanProduccion,
    PlanProduccionItem,
    PronosticoVenta,
    Receta,
    RecetaCodigoPointAlias,
    RecetaCostoVersion,
    RecetaEquivalencia,
    RecetaPresentacion,
    RecetaPresentacionDerivada,
    PoliticaStockSucursalProducto,
    SolicitudReabastoCedis,
    SolicitudReabastoCedisLinea,
    SolicitudVenta,
    VentaHistorica,
)
from pos_bridge.models import PointBranch, PointDailySale, PointSalesDailyProductFact, PointSyncJob
from recetas.utils.costeo_versionado import asegurar_version_costeo, calcular_costeo_receta
from recetas.utils.costeo_snapshot import resolve_line_snapshot_cost
from recetas.utils.derived_insumos import sync_presentacion_insumo, sync_receta_derivados
from recetas.utils.importador import ImportadorCosteo
from pos_bridge.models import PointProduct


class MatchingPendientesAutocompleteTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_match",
            email="admin_match@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

        unidad = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.insumo_1 = Insumo.objects.create(nombre="Harina Pastelera", unidad_base=unidad, activo=True)
        self.insumo_1_canon = Insumo.objects.create(
            nombre="HARINA   PASTELERA",
            codigo_point="PT-HAR-001",
            unidad_base=unidad,
            activo=True,
        )
        self.insumo_2 = Insumo.objects.create(nombre="Harina Integral", unidad_base=unidad, activo=True)
        Insumo.objects.create(nombre="Mantequilla", unidad_base=unidad, activo=True)

        receta = Receta.objects.create(nombre="Receta Test Match", hash_contenido="hash-match-test-001")
        self.linea = LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=None,
            insumo_texto="Harina Pastelera",
            cantidad=Decimal("1"),
            unidad=unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("0"),
            match_status=LineaReceta.STATUS_NEEDS_REVIEW,
            match_score=85,
            match_method="FUZZY",
        )

    def test_matching_pendientes_view_loads(self):
        response = self.client.get(reverse("recetas:matching_pendientes"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Validación BOM de artículos")
        self.assertContains(response, "Resumen del cálculo")
        self.assertContains(response, "Workflow ERP del componente")
        self.assertContains(response, "Ruta crítica ERP")
        self.assertContains(response, "Radar ejecutivo ERP")
        self.assertIn("workflow_rows", response.context)
        self.assertIn("erp_command_center", response.context)
        self.assertIn("critical_path_rows", response.context)

    def test_matching_pendientes_export_csv(self):
        response = self.client.get(reverse("recetas:matching_pendientes"), {"export": "csv", "q": "Harina"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        body = response.content.decode("utf-8")
        self.assertIn("receta,posicion,ingrediente,metodo,score,insumo_ligado", body)
        self.assertIn("Receta Test Match", body)

    def test_matching_pendientes_export_xlsx(self):
        response = self.client.get(reverse("recetas:matching_pendientes"), {"export": "xlsx", "q": "Harina"})
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        wb = load_workbook(BytesIO(response.content), data_only=True)
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, 7)]
        self.assertEqual(headers, ["receta", "posicion", "ingrediente", "metodo", "score", "insumo_ligado"])
        self.assertEqual(ws.cell(row=2, column=1).value, "Receta Test Match")
        self.assertEqual(ws.cell(row=2, column=3).value, "Harina Pastelera")

    def test_matching_pendientes_context_stats(self):
        response = self.client.get(reverse("recetas:matching_pendientes"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mesa de gobierno ERP")
        self.assertIn("erp_governance_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertEqual(response.context["stats"]["total"], 1)
        self.assertEqual(response.context["stats"]["recetas"], 1)
        self.assertEqual(response.context["stats"]["fuzzy"], 1)
        self.assertEqual(response.context["stats"]["no_match"], 0)
        self.assertEqual(response.context["stats"]["auto_suggested"], 1)
        self.assertEqual(response.context["stats"]["canonical_suggested"], 1)

    def test_matching_pendientes_prefers_canonical_suggestion(self):
        response = self.client.get(reverse("recetas:matching_pendientes"))
        self.assertEqual(response.status_code, 200)
        linea = response.context["page"].object_list[0]
        self.assertEqual(linea.suggested_insumo.id, self.insumo_1_canon.id)
        self.assertTrue(linea.suggested_is_canonical)
        self.assertTrue(linea.suggested_can_approve)

    def test_matching_insumos_search_filters_by_query(self):
        response = self.client.get(
            reverse("recetas:matching_insumos_search"),
            {"q": "harina", "limit": "10"},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        names = [x["nombre"] for x in payload["results"]]
        self.assertIn(self.insumo_1_canon.nombre, names)
        self.assertIn(self.insumo_2.nombre, names)
        self.assertNotIn("Mantequilla", names)
        self.assertNotIn(self.insumo_1.nombre, names)

    def test_matching_insumos_search_enforces_limit(self):
        response = self.client.get(
            reverse("recetas:matching_insumos_search"),
            {"q": "harina", "limit": "1"},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["count"], 1)
        self.assertEqual(len(payload["results"]), 1)

    def test_matching_insumos_search_labels_canonical_variants(self):
        response = self.client.get(
            reverse("recetas:matching_insumos_search"),
            {"q": "harina", "limit": "10"},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["results"][0]["id"], self.insumo_1_canon.id)
        self.assertIn("Maestro", payload["results"][0]["label"])
        self.assertEqual(payload["results"][0]["variant_count"], 2)

    def test_aprobar_matching_sugerido_redirects_to_canonical(self):
        response = self.client.post(reverse("recetas:aprobar_matching_sugerido", args=[self.linea.id]))
        self.assertEqual(response.status_code, 302)
        self.linea.refresh_from_db()
        self.assertEqual(self.linea.insumo_id, self.insumo_1_canon.id)
        self.assertEqual(self.linea.match_status, LineaReceta.STATUS_AUTO)
        self.assertIn("CANON", self.linea.match_method)

    def test_aprobar_matching_manual_redirects_to_canonical(self):
        response = self.client.post(
            reverse("recetas:aprobar_matching", args=[self.linea.id]),
            {"insumo_id": str(self.insumo_1.id)},
        )
        self.assertEqual(response.status_code, 302)
        self.linea.refresh_from_db()
        self.assertEqual(self.linea.insumo_id, self.insumo_1_canon.id)
        self.assertEqual(self.linea.match_status, LineaReceta.STATUS_AUTO)
        self.assertEqual(self.linea.match_method, "MANUAL_CANON")

    def test_receta_detail_shows_canonical_suggestion_for_pending_line(self):
        response = self.client.get(reverse("recetas:receta_detail", args=[self.linea.receta.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Resumen del cálculo")
        self.assertContains(response, "Ruta crítica ERP")
        self.assertContains(response, "Sin artículo estándar")
        self.assertContains(response, self.insumo_1_canon.nombre)
        self.assertContains(response, "Cadena de control ERP")
        self.assertContains(response, "Mesa de gobierno ERP")
        self.assertIn("erp_command_center", response.context)
        self.assertIn("erp_governance_rows", response.context)
        self.assertIn("critical_path_rows", response.context)

    def test_aprobar_matching_sugerido_lote_current_page_only(self):
        receta = self.linea.receta
        linea_extra = LineaReceta.objects.create(
            receta=receta,
            posicion=2,
            insumo=None,
            insumo_texto="Harina Integral",
            cantidad=Decimal("2"),
            unidad=self.insumo_2.unidad_base,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("0"),
            match_status=LineaReceta.STATUS_NEEDS_REVIEW,
            match_score=80,
            match_method="FUZZY",
        )
        response = self.client.post(
            reverse("recetas:aprobar_matching_sugerido_lote"),
            {"q": "Harina", "page": "1"},
        )
        self.assertEqual(response.status_code, 302)
        self.linea.refresh_from_db()
        linea_extra.refresh_from_db()
        self.assertEqual(self.linea.insumo_id, self.insumo_1_canon.id)
        self.assertEqual(linea_extra.insumo_id, self.insumo_2.id)
        self.assertEqual(self.linea.match_status, LineaReceta.STATUS_AUTO)
        self.assertEqual(linea_extra.match_status, LineaReceta.STATUS_AUTO)

    def test_aprobar_matching_sugerido_respects_next_redirect(self):
        next_url = reverse("recetas:receta_detail", args=[self.linea.receta.id])
        response = self.client.post(
            reverse("recetas:aprobar_matching_sugerido", args=[self.linea.id]),
            {"next": next_url},
        )
        self.assertRedirects(response, next_url)

    def test_matching_pendientes_can_filter_by_receta(self):
        otra_receta = Receta.objects.create(nombre="Otra receta match", hash_contenido="hash-otra-match-001")
        LineaReceta.objects.create(
            receta=otra_receta,
            posicion=1,
            insumo=None,
            insumo_texto="Mantequilla QA",
            cantidad=Decimal("1.000000"),
            unidad=self.insumo_2.unidad_base,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("0"),
            match_status=LineaReceta.STATUS_NEEDS_REVIEW,
            match_score=70,
            match_method="FUZZY",
        )

        response = self.client.get(reverse("recetas:matching_pendientes"), {"receta": self.linea.receta.id})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.linea.receta.nombre)
        self.assertNotContains(response, otra_receta.nombre)

    def test_aprobar_matching_sugerido_lote_respects_receta_filter(self):
        otra_receta = Receta.objects.create(nombre="Otra receta lote", hash_contenido="hash-otra-match-002")
        linea_otra = LineaReceta.objects.create(
            receta=otra_receta,
            posicion=1,
            insumo=None,
            insumo_texto="Harina Integral",
            cantidad=Decimal("2.000000"),
            unidad=self.insumo_2.unidad_base,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("0"),
            match_status=LineaReceta.STATUS_NEEDS_REVIEW,
            match_score=80,
            match_method="FUZZY",
        )

        response = self.client.post(
            reverse("recetas:aprobar_matching_sugerido_lote"),
            {"receta": str(self.linea.receta.id), "page": "1"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn(f"receta={self.linea.receta.id}", response.url)

        self.linea.refresh_from_db()
        linea_otra.refresh_from_db()
        self.assertEqual(self.linea.insumo_id, self.insumo_1_canon.id)
        self.assertIsNone(linea_otra.insumo_id)

    def test_receta_aprobar_sugeridos_applies_to_current_recipe(self):
        receta = self.linea.receta
        otra_receta = Receta.objects.create(nombre="Otra receta QA", hash_contenido="hash-otra-receta-001")
        otra_linea = LineaReceta.objects.create(
            receta=otra_receta,
            posicion=1,
            insumo=None,
            insumo_texto="Harina Integral",
            cantidad=Decimal("1"),
            unidad=self.insumo_2.unidad_base,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("0"),
            match_status=LineaReceta.STATUS_NEEDS_REVIEW,
            match_score=85,
            match_method="FUZZY",
        )
        response = self.client.post(reverse("recetas:receta_aprobar_sugeridos", args=[receta.id]))
        self.assertRedirects(response, reverse("recetas:receta_detail", args=[receta.id]))
        self.linea.refresh_from_db()
        otra_linea.refresh_from_db()
        self.assertEqual(self.linea.insumo_id, self.insumo_1_canon.id)
        self.assertEqual(self.linea.match_status, LineaReceta.STATUS_AUTO)
        self.assertIsNone(otra_linea.insumo_id)
        self.assertEqual(otra_linea.match_status, LineaReceta.STATUS_NEEDS_REVIEW)

    def test_linea_repoint_canonical_updates_line_to_canonical(self):
        linea = LineaReceta.objects.create(
            receta=self.linea.receta,
            posicion=2,
            insumo=self.insumo_1,
            insumo_texto=self.insumo_1.nombre,
            cantidad=Decimal("1"),
            unidad=self.insumo_1.unidad_base,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("10"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method="MANUAL",
        )
        next_url = reverse("recetas:receta_detail", args=[linea.receta.id])
        response = self.client.post(
            reverse("recetas:linea_repoint_canonical", args=[linea.id]),
            {"next": next_url},
        )
        self.assertRedirects(response, next_url)
        linea.refresh_from_db()
        self.assertEqual(linea.insumo_id, self.insumo_1_canon.id)
        self.assertEqual(linea.insumo_texto, self.insumo_1_canon.nombre)

    def test_receta_repoint_canonical_updates_all_lines_in_recipe(self):
        receta = self.linea.receta
        linea_dup = LineaReceta.objects.create(
            receta=receta,
            posicion=2,
            insumo=self.insumo_1,
            insumo_texto=self.insumo_1.nombre,
            cantidad=Decimal("1"),
            unidad=self.insumo_1.unidad_base,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("10"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method="MANUAL",
        )
        response = self.client.post(reverse("recetas:receta_repoint_canonical", args=[receta.id]))
        self.assertRedirects(response, reverse("recetas:receta_detail", args=[receta.id]))
        linea_dup.refresh_from_db()
        self.assertEqual(linea_dup.insumo_id, self.insumo_1_canon.id)

    def test_receta_detail_shows_canonical_normalization_controls(self):
        LineaReceta.objects.create(
            receta=self.linea.receta,
            posicion=2,
            insumo=self.insumo_1,
            insumo_texto=self.insumo_1.nombre,
            cantidad=Decimal("1"),
            unidad=self.insumo_1.unidad_base,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("10"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method="MANUAL",
        )
        response = self.client.get(reverse("recetas:receta_detail", args=[self.linea.receta.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Artículo propuesto:")
        self.assertContains(response, "Usar artículo ERP")
        self.assertContains(response, "Alinear artículos del maestro")


class RecetasListCatalogFiltersTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_recetas_catalogo",
            email="admin_recetas_catalogo@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)
        self.unidad_kg = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.receta_preparacion = Receta.objects.create(
            nombre="Batida Chocolate",
            hash_contenido="hash-catalogo-001",
            tipo=Receta.TIPO_PREPARACION,
            familia="Batidas",
            categoria="Chocolate",
            rendimiento_cantidad=Decimal("10.000000"),
            rendimiento_unidad=self.unidad_kg,
        )
        self.receta_subinsumo = Receta.objects.create(
            nombre="Batida Vainilla con Presentaciones",
            hash_contenido="hash-catalogo-003",
            tipo=Receta.TIPO_PREPARACION,
            familia="Batidas",
            categoria="Vainilla",
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("12.000000"),
            rendimiento_unidad=self.unidad_kg,
        )
        self.receta_producto = Receta.objects.create(
            nombre="Pastel Fresas con Crema - Chico",
            hash_contenido="hash-catalogo-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pasteles",
            categoria="Frutales",
        )

    def test_recetas_list_default_view_shows_productos(self):
        response = self.client.get(reverse("recetas:recetas_list"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["vista"], "productos")
        self.assertContains(response, "Catálogo de recetas")
        self.assertContains(response, "Prioridad del catálogo")
        self.assertIn("erp_command_center", response.context)
        self.assertIn("erp_governance_rows", response.context)
        self.assertIn("critical_path_rows", response.context)
        nombres = [r.nombre for r in response.context["page"].object_list]
        self.assertIn(self.receta_producto.nombre, nombres)
        self.assertNotIn(self.receta_preparacion.nombre, nombres)

    def test_recetas_list_shows_latest_point_recipe_import_panel(self):
        job = PointSyncJob.objects.create(
            job_type=PointSyncJob.JOB_TYPE_RECIPES,
            status=PointSyncJob.STATUS_SUCCESS,
            parameters={"action": "SYNC_ONLY_NEW_PRODUCTS", "product_codes": ["PT-NUEVO-001"]},
            result_summary={
                "products_selected": 1,
                "recipes_completed_successfully": 1,
                "recipes_with_unresolved_inputs": 0,
                "new_products_imported": 1,
                "new_preparations_imported": 1,
                "recursive_nodes_created": 1,
                "unresolved_inputs_count": 0,
                "point_retry_events": 0,
                "point_relogin_events": 0,
                "imported_products_status": [
                    {
                        "codigo_point": "PT-NUEVO-001",
                        "nombre": "Pastel Nuevo Point",
                        "status": "SUCCESS_COMPLETE",
                        "is_new_product": True,
                        "unresolved_inputs": [],
                        "created_preparations": [
                            {"codigo_point": "PREP-001", "nombre": "Betún Base Point"},
                        ],
                        "message": "Se importó Pastel Nuevo Point correctamente con toda su receta y 1 preparación(es) hija(s) automática(s).",
                    }
                ],
            },
            triggered_by=self.user,
        )
        session = self.client.session
        session["recetas_last_point_sync_job_id"] = job.id
        session.save()

        response = self.client.get(reverse("recetas:recetas_list"))

        self.assertEqual(response.status_code, 200)
        self.assertIn("point_recipe_sync_panel", response.context)
        self.assertEqual(response.context["point_recipe_sync_panel"]["job_id"], job.id)
        self.assertContains(response, "Última importación Point")
        self.assertContains(response, "Pastel Nuevo Point")
        self.assertContains(response, "Producto nuevo importado")
        self.assertContains(response, "Preparaciones creadas automáticamente")

    def test_recetas_list_filters_by_tipo(self):
        response = self.client.get(reverse("recetas:recetas_list"), {"tipo": Receta.TIPO_PRODUCTO_FINAL})
        self.assertEqual(response.status_code, 200)
        nombres = [r.nombre for r in response.context["page"].object_list]
        self.assertIn(self.receta_producto.nombre, nombres)
        self.assertNotIn(self.receta_preparacion.nombre, nombres)

    @patch("recetas.views.PointProductRecipeSyncService")
    def test_recetas_sync_new_warns_when_point_detects_blocked_candidates_without_bom(self, service_cls):
        service = service_cls.return_value
        service.discover_new_product_codes.return_value = {
            "workspace": "Matriz",
            "products_seen": 150,
            "new_candidates": [],
            "new_codes": [],
            "blocked_candidates": [
                {
                    "codigo_point": "EXTRA-FRESA-C",
                    "nombre": "Extra Fresa Chico",
                    "familia": "Otros postres",
                    "categoria": "Pastel Chico",
                    "detection_reason": "POINT_NO_BOM",
                    "message": "Point no reportó receta/BOM para este código y no puede incorporarse automáticamente.",
                    "bom_lines": 0,
                }
            ],
            "blocked_codes": ["EXTRA-FRESA-C"],
            "importable_candidates_count": 0,
            "blocked_candidates_count": 1,
            "ignored_candidates_count": 90,
        }

        response = self.client.post(reverse("recetas:recetas_sync_new"), follow=True)

        self.assertEqual(response.status_code, 200)
        messages = [str(message) for message in response.context["messages"]]
        self.assertTrue(any("bloqueado" in message.lower() for message in messages))
        self.assertTrue(any("EXTRA-FRESA-C" in message for message in messages))

    def test_recetas_list_quick_view_subinsumos(self):
        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "subinsumos", "_debug_chain_focus": "1"})
        self.assertEqual(response.status_code, 200)
        nombres = [r.nombre for r in response.context["page"].object_list]
        self.assertIn(self.receta_subinsumo.nombre, nombres)
        self.assertNotIn(self.receta_preparacion.nombre, nombres)
        self.assertNotIn(self.receta_producto.nombre, nombres)

    def test_recetas_list_filters_by_familia_and_categoria(self):
        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"familia": "Pasteles", "categoria": "Frutales"},
        )
        self.assertEqual(response.status_code, 200)
        page = response.context["page"]
        self.assertEqual(page.paginator.count, 1)
        self.assertEqual(page.object_list[0].id, self.receta_producto.id)
        self.assertIn("Pasteles", response.context["familias_catalogo"])
        self.assertIn("Frutales", response.context["categorias_catalogo"])

    def test_recetas_list_exposes_top_family_and_category_navigation(self):
        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "todo"})
        self.assertEqual(response.status_code, 200)
        top_familias = {item["familia"]: item["total"] for item in response.context["familias_top"]}
        top_categorias = {item["categoria"]: item["total"] for item in response.context["categorias_top"]}
        self.assertEqual(top_familias["Batidas"], 2)
        self.assertEqual(top_familias["Pasteles"], 1)
        self.assertEqual(top_categorias["Chocolate"], 1)
        self.assertEqual(top_categorias["Frutales"], 1)
        self.assertContains(response, "Todas las familias")
        self.assertContains(response, "Todas las categorías")

    def test_recetas_list_filters_by_modo_operativo_base(self):
        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "todo", "modo_operativo": "BASE"},
        )
        self.assertEqual(response.status_code, 200)
        nombres = [r.nombre for r in response.context["page"].object_list]
        self.assertIn(self.receta_preparacion.nombre, nombres)
        self.assertNotIn(self.receta_subinsumo.nombre, nombres)
        self.assertNotIn(self.receta_producto.nombre, nombres)

    def test_receta_create_accepts_mode_from_querystring(self):
        response = self.client.get(reverse("recetas:receta_create"), {"mode": "FINAL"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["values"]["recipe_mode"], "FINAL")

    def test_receta_create_prefills_final_from_source_base(self):
        base = Receta.objects.create(
            nombre="Base Prefill Final",
            hash_contenido="hash-receta-create-source-base-001",
            tipo=Receta.TIPO_PREPARACION,
            familia="Pastel",
            categoria="Chocolate",
            sheet_name="Insumos 1",
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("10.000000"),
            rendimiento_unidad=self.unidad_kg,
        )
        response = self.client.get(
            reverse("recetas:receta_create"),
            {"mode": "FINAL", "source_base": str(base.id)},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["values"]["recipe_mode"], "FINAL")
        self.assertEqual(response.context["values"]["familia"], "Pastel")
        self.assertEqual(response.context["values"]["categoria"], "Chocolate")
        self.assertEqual(response.context["values"]["sheet_name"], "Insumos 1")
        self.assertEqual(response.context["source_base_receta"].id, base.id)
        self.assertIsNotNone(response.context["source_base_context"])
        self.assertEqual(response.context["source_base_context"]["active_presentaciones"], 0)
        self.assertContains(response, "Cierre de cadena operativo desde la base")
        self.assertContains(response, "Presentaciones activas")
        self.assertContains(response, "Derivados activos")

    def test_recetas_list_shows_operational_health(self):
        self.receta_preparacion.rendimiento_cantidad = None
        self.receta_preparacion.save(update_fields=["rendimiento_cantidad"])
        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "todo"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Resumen")
        self.assertContains(response, "Qué falta y qué hacer")
        self.assertContains(response, "Sin rendimiento")
        self.assertContains(response, "Falta rendimiento para costeo enterprise.")

    def test_recetas_list_marks_producto_final_using_base_direct(self):
        base = Receta.objects.create(
            nombre="Base Lista QA",
            hash_contenido="hash-catalogo-base-directa-001",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            familia="Batidas",
            categoria="Chocolate",
            rendimiento_cantidad=Decimal("10.000000"),
            rendimiento_unidad=self.unidad_kg,
        )
        presentacion = RecetaPresentacion.objects.create(
            receta=base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.380000"),
            activo=True,
        )
        Insumo.objects.create(
            nombre="Base QA Directa - Chico",
            codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:{presentacion.id}",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_kg,
            activo=True,
        )
        base_directa = Insumo.objects.create(
            nombre="Base Lista QA",
            codigo=f"DERIVADO:RECETA:{base.id}:PREPARACION",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_kg,
            activo=True,
        )
        LineaReceta.objects.create(
            receta=self.receta_producto,
            posicion=1,
            insumo=base_directa,
            insumo_texto=base_directa.nombre,
            cantidad=Decimal("0.380000"),
            unidad=self.unidad_kg,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("9.000000"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        empaque = Insumo.objects.create(
            nombre="Caja Lista QA",
            tipo_item=Insumo.TIPO_EMPAQUE,
            categoria="Cajas",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        LineaReceta.objects.create(
            receta=self.receta_producto,
            posicion=2,
            insumo=empaque,
            insumo_texto=empaque.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad_kg,
            unidad_texto="pza",
            costo_unitario_snapshot=Decimal("2.000000"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "todo"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Usa base sin presentación")
        self.assertContains(response, "Ver contexto")

    def test_recetas_list_filters_by_health_status(self):
        self.receta_preparacion.rendimiento_cantidad = None
        self.receta_preparacion.save(update_fields=["rendimiento_cantidad"])
        LineaReceta.objects.create(
            receta=self.receta_producto,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_SUBSECCION,
            insumo_texto="Cobertura",
            unidad_texto="kg",
            cantidad=Decimal("0.500000"),
            match_status=LineaReceta.STATUS_AUTO,
            match_method="MANUAL",
            match_score=100,
        )

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "todo", "health_status": "incompletas"},
        )
        self.assertEqual(response.status_code, 200)
        nombres = [r.nombre for r in response.context["page"].object_list]
        self.assertIn(self.receta_preparacion.nombre, nombres)
        self.assertNotIn(self.receta_producto.nombre, nombres)
        self.assertContains(response, "Listas para operar")
        self.assertContains(response, "Pendientes operativos")
        self.assertContains(response, "Incompletas")

    def test_recetas_list_treats_equivalence_as_effective_bom(self):
        parent = Receta.objects.create(
            nombre="Pay Padre QA",
            hash_contenido="hash-catalogo-equivalence-parent",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pay",
            categoria="Pay",
        )
        child = Receta.objects.create(
            nombre="Pay Rebanada QA",
            hash_contenido="hash-catalogo-equivalence-child",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pay",
            categoria="Pay",
            codigo_point="PAY-REB-QA",
            sheet_name="AUTO_POINT_SALES",
        )
        PointProduct.objects.create(sku="PAY-REB-QA", external_id="PAY-REB-QA", name=child.nombre, active=True)
        insumo = Insumo.objects.create(
            nombre="Base Pay QA",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_kg,
            activo=True,
        )
        LineaReceta.objects.create(
            receta=parent,
            posicion=1,
            insumo=insumo,
            insumo_texto=insumo.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad_kg,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("80.000000"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        RecetaEquivalencia.objects.create(
            receta_porcion=child,
            receta_padre=parent,
            factor_conversion=Decimal("8.000000"),
            activo=True,
            fuente="TEST",
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos", "q": child.nombre})

        self.assertEqual(response.status_code, 200)
        result = next(r for r in response.context["page"].object_list if r.nombre == child.nombre)
        self.assertTrue(result.has_equivalence_bom)
        self.assertEqual(result.lineas_count, 0)
        self.assertNotIn("componentes", result.governance_issues)
        self.assertNotEqual(result.operational_health["code"], "danger")
        self.assertEqual(result.costo_efectivo, Decimal("10.000000"))
        self.assertEqual(result.fuente_display, "Equivalencia: Pay Padre QA")
        self.assertContains(response, "$10.00")
        self.assertContains(response, "Equivalencia: Pay Padre QA")
        self.assertNotContains(response, "AUTO_POINT_SALES")

    def test_recetas_list_treats_derived_presentation_as_effective_bom(self):
        parent = Receta.objects.create(
            nombre="Pastel Padre QA",
            hash_contenido="hash-catalogo-derived-parent",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Pastel",
        )
        child = Receta.objects.create(
            nombre="Pastel Rebanada QA",
            hash_contenido="hash-catalogo-derived-child",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Pastel",
            codigo_point="DER-QA-1",
        )
        PointProduct.objects.create(sku="DER-QA-1", external_id="DER-QA-1", name=child.nombre, active=True)
        RecetaPresentacionDerivada.objects.create(
            receta_padre=parent,
            receta_derivada=child,
            codigo_point_derivado="DER-QA-1",
            nombre_derivado=child.nombre,
            unidades_por_padre=Decimal("10.000000"),
            activo=True,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos", "q": child.nombre})

        self.assertEqual(response.status_code, 200)
        result = next(r for r in response.context["page"].object_list if r.nombre == child.nombre)
        self.assertTrue(result.has_derived_bom)
        self.assertEqual(result.lineas_count, 0)
        self.assertNotIn("componentes", result.governance_issues)
        self.assertNotEqual(result.operational_health["code"], "danger")

    def test_recetas_list_default_shows_active_point_products_only(self):
        active = Receta.objects.create(
            nombre="Producto Activo Point QA",
            hash_contenido="hash-catalogo-active-point",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Pastel",
            codigo_point="ACT-POINT-QA",
        )
        archived = Receta.objects.create(
            nombre="Producto Archivado Point QA",
            hash_contenido="hash-catalogo-archived-point",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Pastel",
            codigo_point="ARCH-POINT-QA",
        )
        PointProduct.objects.create(sku="ACT-POINT-QA", external_id="ACT-POINT-QA", name=active.nombre, active=True)
        PointProduct.objects.create(sku="ARCH-POINT-QA", external_id="ARCH-POINT-QA", name=archived.nombre, active=False)

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos"})

        self.assertEqual(response.status_code, 200)
        names = [r.nombre for r in response.context["page"].object_list]
        self.assertIn(active.nombre, names)
        self.assertNotIn(archived.nombre, names)
        self.assertEqual(response.context["recetas_activas_count"], 1)
        self.assertEqual(response.context["recetas_archivadas_count"], 2)

    def test_recetas_list_archivados_tab_shows_without_active_point_product(self):
        active = Receta.objects.create(
            nombre="Producto Activo Tab QA",
            hash_contenido="hash-catalogo-active-tab",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Pastel",
            codigo_point="ACT-TAB-QA",
        )
        archived = Receta.objects.create(
            nombre="Producto Archivado Tab QA",
            hash_contenido="hash-catalogo-archived-tab",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Pastel",
            codigo_point="ARCH-TAB-QA",
        )
        no_point = Receta.objects.create(
            nombre="Producto Sin Point Tab QA",
            hash_contenido="hash-catalogo-no-point-tab",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Pastel",
        )
        PointProduct.objects.create(sku="ACT-TAB-QA", external_id="ACT-TAB-QA", name=active.nombre, active=True)
        PointProduct.objects.create(sku="ARCH-TAB-QA", external_id="ARCH-TAB-QA", name=archived.nombre, active=False)

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "archivados"})

        self.assertEqual(response.status_code, 200)
        names = [r.nombre for r in response.context["page"].object_list]
        self.assertNotIn(active.nombre, names)
        self.assertIn(archived.nombre, names)
        self.assertIn(no_point.nombre, names)

    def test_recetas_list_ignores_products_without_point_signal_in_pending_summary(self):
        interno = Insumo.objects.create(
            nombre="Relleno activo catálogo",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_kg,
            activo=True,
        )
        producto_fuera_point = Receta.objects.create(
            nombre="Producto Fuera Point",
            hash_contenido="hash-catalogo-out-point-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pasteles",
            categoria="Temporal",
        )
        producto_vigente_point = Receta.objects.create(
            nombre="Producto Vigente Point",
            hash_contenido="hash-catalogo-live-point-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pasteles",
            categoria="Temporal",
            codigo_point="LIVE-001",
        )
        PointProduct.objects.create(
            external_id="ext-live-001",
            sku="LIVE-001",
            name=producto_vigente_point.nombre,
            active=True,
        )
        for receta in (producto_fuera_point, producto_vigente_point):
            LineaReceta.objects.create(
                receta=receta,
                posicion=1,
                insumo=interno,
                insumo_texto=interno.nombre,
                cantidad=Decimal("1.000000"),
                unidad=self.unidad_kg,
                unidad_texto="kg",
                costo_unitario_snapshot=Decimal("5.000000"),
                match_status=LineaReceta.STATUS_AUTO,
                match_score=100,
                match_method=LineaReceta.MATCH_EXACT,
            )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["health_summary"]["pendientes"], 1)
        self.assertContains(response, "Fuera de Point")

    def test_recetas_list_pending_filter_excludes_products_without_point_signal(self):
        interno = Insumo.objects.create(
            nombre="Relleno activo filtro",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_kg,
            activo=True,
        )
        producto_fuera_point = Receta.objects.create(
            nombre="Producto Filtro Fuera Point",
            hash_contenido="hash-catalogo-out-point-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pasteles",
            categoria="Temporal",
        )
        producto_vigente_point = Receta.objects.create(
            nombre="Producto Filtro Vigente Point",
            hash_contenido="hash-catalogo-live-point-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pasteles",
            categoria="Temporal",
            codigo_point="LIVE-002",
        )
        PointProduct.objects.create(
            external_id="ext-live-002",
            sku="LIVE-002",
            name=producto_vigente_point.nombre,
            active=True,
        )
        for receta in (producto_fuera_point, producto_vigente_point):
            LineaReceta.objects.create(
                receta=receta,
                posicion=1,
                insumo=interno,
                insumo_texto=interno.nombre,
                cantidad=Decimal("1.000000"),
                unidad=self.unidad_kg,
                unidad_texto="kg",
                costo_unitario_snapshot=Decimal("5.000000"),
                match_status=LineaReceta.STATUS_AUTO,
                match_score=100,
                match_method=LineaReceta.MATCH_EXACT,
            )

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "productos", "health_status": "pendientes"},
        )
        self.assertEqual(response.status_code, 200)
        nombres = [r.nombre for r in response.context["page"].object_list]
        self.assertIn(producto_vigente_point.nombre, nombres)
        self.assertNotIn(producto_fuera_point.nombre, nombres)

    def test_recetas_list_ignores_legacy_placeholders_in_operational_pending(self):
        materia = Insumo.objects.create(
            nombre="Materia lista placeholder",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unidad_kg,
            proveedor_principal=Proveedor.objects.create(nombre="Proveedor Placeholder"),
            activo=True,
        )
        empaque = Insumo.objects.create(
            nombre="Caja lista placeholder",
            tipo_item=Insumo.TIPO_EMPAQUE,
            categoria="Empaque",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        decorado = Insumo.objects.create(
            nombre="Decorado listo placeholder",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unidad_kg,
            proveedor_principal=Proveedor.objects.create(nombre="Proveedor Decorado"),
            activo=True,
        )
        producto = Receta.objects.create(
            nombre="Producto Legacy Placeholder",
            hash_contenido="hash-legacy-placeholder-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pasteles",
            categoria="Especial",
            codigo_point="LEGACY-001",
        )
        PointProduct.objects.create(
            external_id="ext-legacy-001",
            sku="LEGACY-001",
            name=producto.nombre,
            active=True,
        )
        LineaReceta.objects.create(
            receta=producto,
            posicion=1,
            insumo=materia,
            insumo_texto=materia.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad_kg,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5.000000"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        LineaReceta.objects.create(
            receta=producto,
            posicion=2,
            insumo=empaque,
            insumo_texto=empaque.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad_kg,
            unidad_texto="pza",
            costo_unitario_snapshot=Decimal("1.000000"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        LineaReceta.objects.create(
            receta=producto,
            posicion=3,
            insumo=None,
            insumo_texto="Presentación",
            cantidad=None,
            unidad=None,
            unidad_texto="",
            costo_unitario_snapshot=Decimal("0"),
            match_status=LineaReceta.STATUS_REJECTED,
            match_score=63.63,
            match_method=LineaReceta.MATCH_NONE,
        )
        LineaReceta.objects.create(
            receta=producto,
            posicion=4,
            insumo=None,
            insumo_texto="Armado",
            cantidad=None,
            unidad=None,
            unidad_texto="",
            costo_unitario_snapshot=Decimal("0"),
            match_status=LineaReceta.STATUS_REJECTED,
            match_score=71.42,
            match_method=LineaReceta.MATCH_NONE,
        )
        LineaReceta.objects.create(
            receta=producto,
            posicion=5,
            tipo_linea=LineaReceta.TIPO_SUBSECCION,
            insumo=decorado,
            insumo_texto="Decorado",
            cantidad=Decimal("0.050000"),
            unidad=self.unidad_kg,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("2.000000"),
            match_status=LineaReceta.STATUS_NEEDS_REVIEW,
            match_score=80,
            match_method=LineaReceta.MATCH_FUZZY,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos"})
        self.assertEqual(response.status_code, 200)
        health = {r.nombre: r.operational_health["label"] for r in response.context["page"].object_list}
        self.assertEqual(health[producto.nombre], "Lista para operar")

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "productos", "estado": "pendientes"},
        )
        self.assertEqual(response.status_code, 200)
        nombres = [r.nombre for r in response.context["page"].object_list]
        self.assertNotIn(producto.nombre, nombres)

    def test_recetas_list_does_not_require_bom_for_resale_product(self):
        reventa = Receta.objects.create(
            nombre="Té de reventa QA",
            hash_contenido="hash-catalogo-reventa-sin-bom",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            modo_costeo=Receta.MODO_COSTEO_REVENTA,
            codigo_point="REVENTA-QA",
            sheet_name="AUTO_POINT_SALES",
        )
        PointProduct.objects.create(
            sku="REVENTA-QA",
            external_id="REVENTA-QA",
            name=reventa.nombre,
            active=True,
        )

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "productos", "estado": "pendientes"},
        )

        self.assertEqual(response.status_code, 200)
        nombres = [r.nombre for r in response.context["page"].object_list]
        self.assertNotIn(reventa.nombre, nombres)

    def test_recetas_list_filters_by_governance_issue(self):
        self.receta_preparacion.rendimiento_cantidad = None
        self.receta_preparacion.save(update_fields=["rendimiento_cantidad"])
        self.receta_producto.familia = ""
        self.receta_producto.save(update_fields=["familia"])

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "todo", "governance_issue": "rendimiento"},
        )
        self.assertEqual(response.status_code, 200)
        nombres = [r.nombre for r in response.context["page"].object_list]
        self.assertIn(self.receta_preparacion.nombre, nombres)
        self.assertNotIn(self.receta_producto.nombre, nombres)

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "todo", "governance_issue": "familia"},
        )
        self.assertEqual(response.status_code, 200)
        nombres = [r.nombre for r in response.context["page"].object_list]
        self.assertIn(self.receta_producto.nombre, nombres)
        self.assertContains(response, "Sin familia")
        self.assertContains(response, "Sin componentes")

    def test_recetas_list_filters_by_governance_maestro_incompleto(self):
        unidad = UnidadMedida.objects.create(
            codigo="kg-gov",
            nombre="Kilogramo GOV",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        insumo_incompleto = Insumo.objects.create(
            nombre="Insumo Maestro Incompleto",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=unidad,
            activo=True,
            categoria="",
        )
        receta = Receta.objects.create(
            nombre="Producto Maestro Incompleto",
            hash_contenido="hash-maestro-incompleto-receta",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=insumo_incompleto,
            insumo_texto=insumo_incompleto.nombre,
            cantidad=Decimal("1.000000"),
            unidad=unidad,
            unidad_texto=unidad.codigo,
            costo_unitario_snapshot=Decimal("5.00"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "todo", "governance_issue": "maestro_incompleto"},
        )
        self.assertEqual(response.status_code, 200)
        nombres = [r.nombre for r in response.context["page"].object_list]
        self.assertIn(receta.nombre, nombres)
        self.assertContains(response, "Faltante dominante:")
        self.assertContains(response, "categoría")
        self.assertContains(response, "(1)")
        self.assertContains(response, "Qué falta y qué hacer")

    def test_recetas_list_primary_action_for_maestro_incompleto_uses_dominant_missing_field(self):
        unidad = UnidadMedida.objects.create(
            codigo="kg-gov-action",
            nombre="Kilogramo GOV Action",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        insumo_incompleto = Insumo.objects.create(
            nombre="Insumo Maestro Incompleto Acción",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=unidad,
            activo=True,
            categoria="Harinas",
        )
        receta = Receta.objects.create(
            nombre="Base Maestro Acción",
            hash_contenido="hash-maestro-incompleto-accion-receta",
            tipo=Receta.TIPO_PREPARACION,
            familia="Bases",
            categoria="Chocolate",
            rendimiento_cantidad=Decimal("5.000000"),
            rendimiento_unidad=unidad,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=insumo_incompleto,
            insumo_texto=insumo_incompleto.nombre,
            cantidad=Decimal("1.000000"),
            unidad=unidad,
            unidad_texto=unidad.codigo,
            costo_unitario_snapshot=Decimal("5.00"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "todo", "q": receta.nombre})

        self.assertEqual(response.status_code, 200)
        receta_context = next(r for r in response.context["page"].object_list if r.id == receta.id)
        self.assertIn("missing_field=proveedor", receta_context.primary_action["url"])
        self.assertIn("enterprise_status=incompletos", receta_context.primary_action["url"])
        self.assertIn("usage_scope=recipes", receta_context.primary_action["url"])

    def test_recetas_list_filters_by_sync_derivados_issue(self):
        RecetaPresentacion.objects.bulk_create(
            [
                RecetaPresentacion(
                    receta=self.receta_subinsumo,
                    nombre="Chico",
                    peso_por_unidad_kg=Decimal("0.380000"),
                    activo=True,
                )
            ]
        )

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "todo", "governance_issue": "sync_derivados"},
        )
        self.assertEqual(response.status_code, 200)
        nombres = [r.nombre for r in response.context["page"].object_list]
        self.assertIn(self.receta_subinsumo.nombre, nombres)
        self.assertContains(response, "Sincronizar derivados")

    def test_recetas_list_shows_derived_state_summary_for_preparaciones(self):
        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "todo"})
        self.assertEqual(response.status_code, 200)
        receta = next(r for r in response.context["page"].object_list if r.id == self.receta_subinsumo.id)
        self.assertIsNotNone(receta.derived_state)
        self.assertGreaterEqual(receta.derived_state["active_presentaciones"], 0)


class RecetaDerivedInsumoAutolinkTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_receta_derivados",
            email="admin_receta_derivados@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

        self.unidad_pza = UnidadMedida.objects.create(
            codigo="pza",
            nombre="Pieza",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        self.unidad_kg = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )

    def test_linea_form_producto_final_uses_quick_mode(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Modo Rapido",
            hash_contenido="hash-quick-mode-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        response = self.client.get(reverse("recetas:linea_create", args=[receta.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Resumen operativo")
        self.assertContains(response, "Ver datos calculados del artículo")
        self.assertContains(response, 'name="modo_rapido"')
        self.assertContains(response, "Guardar componente")
        self.assertContains(response, "Artículo seleccionado")
        self.assertNotContains(response, 'label for="insumo_texto"')

    def test_linea_form_preparacion_uses_full_mode(self):
        receta = Receta.objects.create(
            nombre="Batida QA Modo Completo",
            hash_contenido="hash-quick-mode-002",
            tipo=Receta.TIPO_PREPARACION,
        )
        response = self.client.get(reverse("recetas:linea_create", args=[receta.id]))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'name="modo_rapido"')
        self.assertContains(response, "Tipo de línea")
        self.assertContains(response, "Captura complementaria y estructura opcional")
        self.assertContains(response, "Opciones de estructura (subsección)")

    def test_linea_form_producto_final_can_switch_to_advanced_mode(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Avanzado",
            hash_contenido="hash-quick-mode-003",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        response = self.client.get(reverse("recetas:linea_create", args=[receta.id]), {"advanced": "1"})
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'name="modo_rapido"')
        self.assertContains(response, 'name="advanced_mode" value="1"')
        self.assertContains(response, "Tipo de línea")

    def test_linea_form_producto_final_prefills_component_filter_from_querystring(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Filtro",
            hash_contenido="hash-quick-mode-004",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        response = self.client.get(
            reverse("recetas:linea_create", args=[receta.id]),
            {"component_kind": "EMPAQUE"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["default_component_filter"], "EMPAQUE")
        self.assertContains(response, 'name="component_filter"')
        self.assertContains(response, "Busca caja, domo, etiqueta, vaso o accesorio final")
        self.assertContains(response, "Agregar empaque")
        self.assertContains(response, "Crear artículo")

    def test_linea_form_producto_final_prefills_component_context_from_querystring(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Contexto",
            hash_contenido="hash-quick-mode-004ca",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        response = self.client.get(
            reverse("recetas:linea_create", args=[receta.id]),
            {"component_kind": "EMPAQUE", "component_context": "empaques"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Bloque:")
        self.assertContains(response, 'name="component_context" value="empaques"')

    def test_linea_form_uses_canonical_option_when_duplicate_exists(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Selector Canon",
            hash_contenido="hash-quick-mode-004b",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        Insumo.objects.create(
            nombre="Etiqueta Canon Test",
            unidad_base=self.unidad_pza,
            activo=True,
        )
        canon = Insumo.objects.create(
            nombre="ETIQUETA CANON TEST",
            codigo_point="PT-ETQ-001",
            unidad_base=self.unidad_pza,
            activo=True,
        )

        response = self.client.get(reverse("recetas:linea_create", args=[receta.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, canon.nombre)
        option_ids = [insumo.id for insumo in response.context["insumos"] if "ETIQUETA CANON TEST" in insumo.nombre.upper()]
        self.assertEqual(option_ids, [canon.id])
        self.assertNotContains(response, '>Etiqueta Canon Test [')

    def test_linea_form_producto_final_prioritizes_recipe_context_internal_items(self):
        receta = Receta.objects.create(
            nombre="Pastel Chocolate Especial",
            hash_contenido="hash-quick-mode-004bb",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pasteles",
            categoria="Chocolate",
        )
        vainilla = Insumo.objects.create(
            nombre="Pan Vainilla Dawn",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_kg,
            activo=True,
        )
        chocolate = Insumo.objects.create(
            nombre="Pan Chocolate Dawn",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_kg,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=vainilla,
            fecha=timezone.localdate(),
            moneda="MXN",
            costo_unitario=Decimal("35.000000"),
            source_hash="test-vainilla-context-001",
            raw={},
        )
        CostoInsumo.objects.create(
            insumo=chocolate,
            fecha=timezone.localdate(),
            moneda="MXN",
            costo_unitario=Decimal("40.000000"),
            source_hash="test-chocolate-context-001",
            raw={},
        )

        response = self.client.get(reverse("recetas:linea_create", args=[receta.id]))
        self.assertEqual(response.status_code, 200)
        internos = list(response.context["insumos_internos"])
        nombres = [item.nombre for item in internos[:2]]
        self.assertEqual(nombres[0], "Pan Chocolate Dawn")
        self.assertIn("Pan Vainilla Dawn", nombres)

    def test_linea_form_producto_final_prioritizes_presentacion_over_base_direct(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Presentacion Preferida",
            hash_contenido="hash-quick-mode-004bbc",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        base = Receta.objects.create(
            nombre="Pan Chocolate Base QA",
            hash_contenido="hash-quick-mode-004bbd",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("5"),
            rendimiento_unidad=self.unidad_kg,
        )
        RecetaPresentacion.objects.create(
            receta=base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.380000"),
            activo=True,
        )
        base_directa = Insumo.objects.create(
            nombre="Pan Chocolate Base QA",
            codigo=f"DERIVADO:RECETA:{base.id}:PREPARACION",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_kg,
            activo=True,
        )
        derivado = Insumo.objects.create(
            nombre="Pan Chocolate Base QA - Chico",
            codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:1",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_pza,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=base_directa,
            fecha=timezone.localdate(),
            moneda="MXN",
            costo_unitario=Decimal("18.000000"),
            source_hash="test-base-direct-cost-001",
            raw={},
        )
        CostoInsumo.objects.create(
            insumo=derivado,
            fecha=timezone.localdate(),
            moneda="MXN",
            costo_unitario=Decimal("16.560000"),
            source_hash="test-derived-cost-001",
            raw={},
        )

        response = self.client.get(reverse("recetas:linea_create", args=[receta.id]))
        self.assertEqual(response.status_code, 200)
        internos = [item for item in response.context["insumos_internos"] if "Pan Chocolate Base QA" in item.nombre]
        self.assertGreaterEqual(len(internos), 2)
        self.assertEqual(internos[0].nombre, "Pan Chocolate Base QA - Chico")
        self.assertEqual(internos[1].nombre, "Pan Chocolate Base QA")

    def test_linea_form_producto_final_defaults_to_internal_guidance(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Internos",
            hash_contenido="hash-quick-mode-004c",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        response = self.client.get(reverse("recetas:linea_create", args=[receta.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Busca pan, relleno, cobertura, batida o subinsumo")
        self.assertContains(response, "Crear artículo")
        self.assertContains(response, "Agregar")
        self.assertContains(response, 'id="selection_quality_notice"')
        self.assertNotContains(response, 'data-component-filter="ALL"')

    def test_linea_create_producto_final_quick_mode_requires_canonical_insumo(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Requiere Canonico",
            hash_contenido="hash-quick-mode-004d",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        response = self.client.post(
            reverse("recetas:linea_create", args=[receta.id]),
            data={
                "component_filter": "INSUMO_INTERNO",
                "cantidad": "1.00",
                "insumo_texto": "Pan vainilla chico",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(LineaReceta.objects.filter(receta=receta).count(), 0)

    def test_linea_create_producto_final_quick_mode_requires_cost_vigente(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Sin Costo",
            hash_contenido="hash-quick-mode-004e",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        pan = Insumo.objects.create(
            nombre="Pan QA Sin Costo",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_kg,
            activo=True,
        )
        response = self.client.post(
            reverse("recetas:linea_create", args=[receta.id]),
            data={
                "component_filter": "INSUMO_INTERNO",
                "insumo_id": str(pan.id),
                "cantidad": "0.500000",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "no tiene costo vigente")
        self.assertEqual(LineaReceta.objects.filter(receta=receta).count(), 0)

    def test_linea_pan_autolink_to_derived_presentacion(self):
        receta = Receta.objects.create(
            nombre="Pastel 3 Pecados - Chico",
            hash_contenido="hash-autolink-pan-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        pan_generico = Insumo.objects.create(
            nombre="Pan Chocolate",
            unidad_base=self.unidad_pza,
            activo=True,
        )
        pan_derivado = Insumo.objects.create(
            codigo="DERIVADO:RECETA:114:PRESENTACION:24",
            nombre="Pan de Chocolate Deleite Dawn - Chico",
            unidad_base=self.unidad_pza,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=pan_derivado,
            fecha=timezone.localdate(),
            moneda="MXN",
            costo_unitario=Decimal("16.560000"),
            source_hash="test-pan-derivado-cost-001",
            raw={},
        )

        response = self.client.post(
            reverse("recetas:linea_create", args=[receta.id]),
            data={
                "posicion": "1",
                "tipo_linea": LineaReceta.TIPO_NORMAL,
                "etapa": "",
                "insumo_texto": "Pan Chocolate",
                "insumo_id": str(pan_generico.id),
                "cantidad": "2",
                "unidad_id": str(self.unidad_pza.id),
                "unidad_texto": "pza",
            },
        )
        self.assertEqual(response.status_code, 302)

        linea = LineaReceta.objects.get(receta=receta, posicion=1)
        self.assertEqual(linea.insumo_id, pan_derivado.id)
        self.assertEqual(linea.match_status, LineaReceta.STATUS_AUTO)

    def test_linea_bollo_autolink_prefers_preparacion_kg(self):
        receta = Receta.objects.create(
            nombre="Bollo Chocolate",
            hash_contenido="hash-autolink-bollo-pan-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        pan_generico = Insumo.objects.create(
            nombre="Pan Chocolate",
            unidad_base=self.unidad_pza,
            activo=True,
        )
        pan_preparacion = Insumo.objects.create(
            codigo="DERIVADO:RECETA:29:PREPARACION",
            nombre="Pan de Chocolate Deleite Dawn",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        pan_bollo = Insumo.objects.create(
            codigo="DERIVADO:RECETA:29:PRESENTACION:99",
            nombre="Pan de Chocolate Deleite Dawn - Bollos",
            unidad_base=self.unidad_pza,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=pan_preparacion,
            fecha=timezone.localdate(),
            moneda="MXN",
            costo_unitario=Decimal("42.402733"),
            source_hash="test-bollo-prep-cost-001",
            raw={},
        )
        CostoInsumo.objects.create(
            insumo=pan_bollo,
            fecha=timezone.localdate(),
            moneda="MXN",
            costo_unitario=Decimal("3.180205"),
            source_hash="test-bollo-pres-cost-001",
            raw={},
        )

        response = self.client.post(
            reverse("recetas:linea_create", args=[receta.id]),
            data={
                "tipo_linea": LineaReceta.TIPO_NORMAL,
                "etapa": "",
                "insumo_texto": "Pan Chocolate",
                "insumo_id": str(pan_generico.id),
                "cantidad": "0.075",
            },
        )
        self.assertEqual(response.status_code, 302)

        linea = LineaReceta.objects.get(receta=receta, posicion=1)
        self.assertEqual(linea.insumo_id, pan_preparacion.id)
        self.assertEqual(linea.unidad_id, self.unidad_kg.id)
        self.assertEqual(linea.costo_unitario_snapshot, Decimal("42.402733"))
        self.assertAlmostEqual(linea.costo_total_estimado or 0, 3.180204975, places=6)

    def test_linea_bollo_edit_refreshes_snapshot_when_autolink_changes_insumo(self):
        receta = Receta.objects.create(
            nombre="Bollo Chocolate",
            hash_contenido="hash-autolink-bollo-pan-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        pan_generico = Insumo.objects.create(
            nombre="Pan Chocolate",
            unidad_base=self.unidad_pza,
            activo=True,
        )
        pan_preparacion = Insumo.objects.create(
            codigo="DERIVADO:RECETA:29:PREPARACION",
            nombre="Pan de Chocolate Deleite Dawn",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        pan_bollo = Insumo.objects.create(
            codigo="DERIVADO:RECETA:29:PRESENTACION:98",
            nombre="Pan de Chocolate Deleite Dawn - Bollos",
            unidad_base=self.unidad_pza,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=pan_preparacion,
            fecha=timezone.localdate(),
            moneda="MXN",
            costo_unitario=Decimal("42.402733"),
            source_hash="test-bollo-prep-cost-002",
            raw={},
        )
        CostoInsumo.objects.create(
            insumo=pan_bollo,
            fecha=timezone.localdate(),
            moneda="MXN",
            costo_unitario=Decimal("3.180205"),
            source_hash="test-bollo-pres-cost-002",
            raw={},
        )

        linea = LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=pan_bollo,
            insumo_texto="Pan Chocolate",
            cantidad=Decimal("0.075000"),
            unidad=self.unidad_pza,
            unidad_texto="pza",
            costo_unitario_snapshot=Decimal("3.180205"),
            match_status=LineaReceta.STATUS_AUTO,
            match_method="MANUAL",
            match_score=100.0,
        )

        response = self.client.post(
            reverse("recetas:linea_edit", args=[receta.id, linea.id]),
            data={
                "tipo_linea": LineaReceta.TIPO_NORMAL,
                "insumo_texto": "Pan Chocolate",
                "insumo_id": str(pan_generico.id),
                "cantidad": "0.075",
                "etapa": "",
            },
        )
        self.assertEqual(response.status_code, 302)
        linea.refresh_from_db()

        self.assertEqual(linea.insumo_id, pan_preparacion.id)
        self.assertEqual(linea.unidad_id, self.unidad_kg.id)
        self.assertEqual(linea.costo_unitario_snapshot, Decimal("42.402733"))
        self.assertAlmostEqual(linea.costo_total_estimado or 0, 3.180204975, places=6)

    def test_linea_pan_autolink_does_not_fall_back_to_unrelated_flavor(self):
        receta = Receta.objects.create(
            nombre="Bollo Red Velvet",
            hash_contenido="hash-autolink-bollo-red-velvet-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        pan_generico = Insumo.objects.create(
            nombre="Pan Red Velvet",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=pan_generico,
            fecha=timezone.localdate(),
            moneda="MXN",
            costo_unitario=Decimal("12.000000"),
            source_hash="test-bollo-red-velvet-base-cost-001",
            raw={},
        )
        pan_tres_leches = Insumo.objects.create(
            codigo="DERIVADO:RECETA:77:PREPARACION",
            nombre="Pan 3 Leches",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=pan_tres_leches,
            fecha=timezone.localdate(),
            moneda="MXN",
            costo_unitario=Decimal("10.000000"),
            source_hash="test-bollo-red-velvet-no-fallback-001",
            raw={},
        )

        response = self.client.post(
            reverse("recetas:linea_create", args=[receta.id]),
            data={
                "tipo_linea": LineaReceta.TIPO_NORMAL,
                "etapa": "",
                "insumo_texto": "Pan Red Velvet",
                "insumo_id": str(pan_generico.id),
                "cantidad": "0.090000",
            },
        )
        self.assertEqual(response.status_code, 302)

        linea = LineaReceta.objects.get(receta=receta, posicion=1)
        self.assertEqual(linea.insumo_id, pan_generico.id)

    def test_signals_sync_prepare_and_presentacion_derived_insumos(self):
        receta = Receta.objects.create(
            nombre="Batida Test Chocolate",
            hash_contenido="hash-signal-derived-001",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("10.000000"),
            rendimiento_unidad=self.unidad_kg,
        )
        prep_code = f"DERIVADO:RECETA:{receta.id}:PREPARACION"
        self.assertTrue(Insumo.objects.filter(codigo=prep_code, activo=True).exists())

        presentacion = RecetaPresentacion.objects.create(
            receta=receta,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.280000"),
            activo=True,
        )
        pres_code = f"DERIVADO:RECETA:{receta.id}:PRESENTACION:{presentacion.id}"
        self.assertTrue(Insumo.objects.filter(codigo=pres_code, activo=True).exists())

    def test_sync_receta_derivados_keeps_preparacion_active_when_downstream_usage_exists(self):
        base_final = Receta.objects.create(
            nombre="Galleta Lotus QA",
            hash_contenido="hash-signal-derived-downstream-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("1.000000"),
            rendimiento_unidad=self.unidad_kg,
        )
        prep_code = f"DERIVADO:RECETA:{base_final.id}:PREPARACION"
        prep = Insumo.objects.create(
            codigo=prep_code,
            nombre=base_final.nombre,
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_kg,
            activo=True,
        )
        presentacion = RecetaPresentacion.objects.create(
            receta=base_final,
            nombre="Individual",
            peso_por_unidad_kg=Decimal("0.090000"),
            activo=True,
        )
        producto_final = Receta.objects.create(
            nombre="Pastel QA usa base producto final",
            hash_contenido="hash-signal-derived-downstream-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=producto_final,
            posicion=1,
            insumo=prep,
            insumo_texto=prep.nombre,
            cantidad=Decimal("0.250000"),
            unidad=self.unidad_kg,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        sync_receta_derivados(base_final)

        prep.refresh_from_db()
        self.assertTrue(prep.activo)

    def test_producto_final_requires_linked_insumo_in_main_lines(self):
        receta = Receta.objects.create(
            nombre="Pastel QA",
            hash_contenido="hash-qa-prod-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        response = self.client.post(
            reverse("recetas:linea_create", args=[receta.id]),
            data={
                "posicion": "1",
                "tipo_linea": LineaReceta.TIPO_NORMAL,
                "insumo_texto": "Armado libre",
                "cantidad": "1",
                "unidad_id": str(self.unidad_pza.id),
                "unidad_texto": "pza",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(LineaReceta.objects.filter(receta=receta).exists())

    def test_linked_insumo_requires_positive_qty(self):
        receta = Receta.objects.create(
            nombre="Pastel QA 2",
            hash_contenido="hash-qa-prod-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        pan = Insumo.objects.create(
            nombre="Pan Vainilla QA",
            unidad_base=self.unidad_pza,
            activo=True,
        )
        response = self.client.post(
            reverse("recetas:linea_create", args=[receta.id]),
            data={
                "posicion": "1",
                "tipo_linea": LineaReceta.TIPO_NORMAL,
                "insumo_texto": "Pan Vainilla QA",
                "insumo_id": str(pan.id),
                "cantidad": "",
                "unidad_id": str(self.unidad_kg.id),
                "unidad_texto": "kg",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(LineaReceta.objects.filter(receta=receta).exists())

    def test_quick_mode_persists_canonical_name_from_selected_insumo(self):
        receta = Receta.objects.create(
            nombre="Pastel QA 3",
            hash_contenido="hash-qa-prod-003",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        pan = Insumo.objects.create(
            nombre="Pan Vainilla Canonico",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=pan,
            fecha=timezone.localdate(),
            moneda="MXN",
            costo_unitario=Decimal("44.500000"),
            source_hash="test-pan-canon-cost-001",
            raw={},
        )
        response = self.client.post(
            reverse("recetas:linea_create", args=[receta.id]),
            data={
                "insumo_texto": "texto libre incorrecto",
                "insumo_id": str(pan.id),
                "cantidad": "0.500000",
                "unidad_texto": "otra",
            },
        )
        self.assertEqual(response.status_code, 302)
        linea = LineaReceta.objects.get(receta=receta)
        self.assertEqual(linea.insumo_id, pan.id)
        self.assertEqual(linea.insumo_texto, "Pan Vainilla Canonico")
        self.assertEqual(linea.unidad_id, self.unidad_kg.id)

    def test_linea_create_auto_repoints_selected_duplicate_to_canonical(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Canon",
            hash_contenido="hash-qa-prod-005",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        harina_dup = Insumo.objects.create(
            nombre="Harina Canon Test",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        harina_canon = Insumo.objects.create(
            nombre="HARINA CANON TEST",
            codigo_point="PT-CAN-001",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=harina_canon,
            fecha=timezone.localdate(),
            moneda="MXN",
            costo_unitario=Decimal("23.400000"),
            source_hash="test-harina-canon-cost-001",
            raw={},
        )
        response = self.client.post(
            reverse("recetas:linea_create", args=[receta.id]),
            data={
                "component_filter": "MATERIA_PRIMA",
                "insumo_texto": harina_dup.nombre,
                "insumo_id": str(harina_dup.id),
                "cantidad": "1.000000",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        linea = LineaReceta.objects.get(receta=receta)
        self.assertEqual(linea.insumo_id, harina_canon.id)
        self.assertEqual(linea.insumo_texto, harina_canon.nombre)
        self.assertContains(response, "se normalizó automáticamente")

    def test_component_filter_rejects_wrong_item_class(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Empaque",
            hash_contenido="hash-qa-prod-004",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        mp = Insumo.objects.create(
            nombre="Fresa QA Filter",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unidad_kg,
            activo=True,
        )
        response = self.client.post(
            reverse("recetas:linea_create", args=[receta.id]),
            data={
                "component_filter": "EMPAQUE",
                "insumo_texto": "Fresa QA Filter",
                "insumo_id": str(mp.id),
                "cantidad": "0.100000",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Seleccionaste el flujo de empaque")
        self.assertFalse(LineaReceta.objects.filter(receta=receta).exists())

    def test_component_filter_rejects_internal_item_missing_categoria(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Interno Incompleto",
            hash_contenido="hash-qa-prod-006",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        interno = Insumo.objects.create(
            nombre="Relleno QA Sin Categoria",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_kg,
            categoria="",
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=interno,
            fecha=timezone.localdate(),
            moneda="MXN",
            costo_unitario=Decimal("15.000000"),
            source_hash="test-interno-no-categoria-cost-001",
            raw={},
        )

        response = self.client.post(
            reverse("recetas:linea_create", args=[receta.id]),
            data={
                "component_filter": "INSUMO_INTERNO",
                "insumo_texto": interno.nombre,
                "insumo_id": str(interno.id),
                "cantidad": "0.500000",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "no está listo para operar en ERP")
        self.assertContains(response, "categoría")
        self.assertFalse(LineaReceta.objects.filter(receta=receta).exists())

    def test_linea_form_prioritizes_erp_ready_internal_items(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Priorizacion",
            hash_contenido="hash-qa-prod-007",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        interno_incompleto = Insumo.objects.create(
            nombre="Chocolate QA Incompleto",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_kg,
            categoria="",
            activo=True,
        )
        interno_listo = Insumo.objects.create(
            nombre="Chocolate QA Listo",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_kg,
            categoria="Batidas",
            activo=True,
        )
        response = self.client.get(reverse("recetas:linea_create", args=[receta.id]))

        self.assertEqual(response.status_code, 200)
        internos = response.context["insumos_internos"]
        ready_idx = next(i for i, item in enumerate(internos) if item.id == interno_listo.id)
        incomplete_idx = next(i for i, item in enumerate(internos) if item.id == interno_incompleto.id)
        self.assertLess(ready_idx, incomplete_idx)
        self.assertContains(response, "Mostrar artículos incompletos")
        self.assertEqual(response.context["insumos_internos_ready_count"], 1)


class RecetaPresentacionCosteoTests(TestCase):
    def setUp(self):
        self.unidad_kg = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.unidad_lt = UnidadMedida.objects.create(
            codigo="lt",
            nombre="Litro",
            tipo=UnidadMedida.TIPO_VOLUMEN,
            factor_to_base=Decimal("1000"),
        )

    def test_costo_presentacion_calcula_para_receta_en_kg(self):
        insumo = Insumo.objects.create(nombre="Harina Test KG", unidad_base=self.unidad_kg, activo=True)
        receta = Receta.objects.create(
            nombre="Batida KG",
            hash_contenido="hash-pres-kg-001",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("10.000000"),
            rendimiento_unidad=self.unidad_kg,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=insumo,
            insumo_texto="Harina Test KG",
            cantidad=Decimal("2.000000"),
            unidad=self.unidad_kg,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5.000000"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        presentacion = RecetaPresentacion.objects.create(
            receta=receta,
            nombre="Mini",
            peso_por_unidad_kg=Decimal("0.500000"),
            activo=True,
        )
        self.assertEqual(presentacion.costo_por_unidad_estimado, Decimal("0.500000"))

    def test_costo_presentacion_calcula_para_receta_en_lt(self):
        insumo = Insumo.objects.create(nombre="Leche Test LT", unidad_base=self.unidad_lt, activo=True)
        receta = Receta.objects.create(
            nombre="Flan Base LT",
            hash_contenido="hash-pres-lt-001",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("20.000000"),
            rendimiento_unidad=self.unidad_lt,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=insumo,
            insumo_texto="Leche Test LT",
            cantidad=Decimal("4.000000"),
            unidad=self.unidad_lt,
            unidad_texto="lt",
            costo_unitario_snapshot=Decimal("25.000000"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        presentacion = RecetaPresentacion.objects.create(
            receta=receta,
            nombre="Vaso",
            peso_por_unidad_kg=Decimal("0.750000"),
            activo=True,
        )
        self.assertEqual(presentacion.costo_por_unidad_estimado, Decimal("3.750000"))


class RecetaPresentacionWorkflowTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_receta_presentaciones",
            email="admin_receta_presentaciones@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)
        self.unidad_kg = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )

    def test_presentacion_create_rejects_producto_final(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Presentaciones",
            hash_contenido="hash-presentacion-producto-final-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pasteles",
        )

        response = self.client.get(reverse("recetas:presentacion_create", args=[receta.id]), follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Solo un insumo base permite presentaciones derivadas.")

    def test_presentacion_create_rejects_base_without_rendimiento(self):
        receta = Receta.objects.create(
            nombre="Batida QA Sin Rendimiento Presentaciones",
            hash_contenido="hash-presentacion-sin-rendimiento-001",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
        )

        response = self.client.get(reverse("recetas:presentacion_create", args=[receta.id]), follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Antes de administrar presentaciones debes capturar el rendimiento total de la base.")
        self.assertContains(response, "Antes de administrar presentaciones debes capturar la unidad del rendimiento.")

    def test_presentacion_create_renders_enterprise_cockpit(self):
        receta = Receta.objects.create(
            nombre="Batida QA Presentacion ERP",
            hash_contenido="hash-presentacion-erp-001",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("10.000000"),
            rendimiento_unidad=self.unidad_kg,
        )

        response = self.client.get(reverse("recetas:presentacion_create", args=[receta.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Resumen del cálculo")
        self.assertContains(response, "Workflow ERP del derivado")
        self.assertContains(response, "Nueva presentación")

    def test_receta_detail_shows_presentacion_health_for_base_derivados(self):
        receta = Receta.objects.create(
            nombre="Pan QA Estado Derivados",
            hash_contenido="hash-presentacion-health-001",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("10.400000"),
            rendimiento_unidad=self.unidad_kg,
        )
        insumo = Insumo.objects.create(
            nombre="Harina QA Estado Derivados",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unidad_kg,
            activo=True,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=insumo,
            insumo_texto=insumo.nombre,
            unidad=self.unidad_kg,
            unidad_texto="kg",
            cantidad=Decimal("2.000000"),
            costo_unitario_snapshot=Decimal("5.000000"),
            match_status=LineaReceta.STATUS_AUTO,
            match_method="MANUAL",
            match_score=100,
        )
        presentacion = RecetaPresentacion.objects.create(
            receta=receta,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.380000"),
            activo=True,
        )
        sync_presentacion_insumo(presentacion)

        response = self.client.get(reverse("recetas:receta_detail", args=[receta.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Estado de derivados")
        self.assertContains(response, "Listo para derivados")
        self.assertContains(response, "Presentaciones activas")
        self.assertContains(response, "Derivados activos")
        self.assertContains(response, "Administrar presentaciones")

    def test_receta_detail_shows_chain_actions_for_pending_derivados_and_no_final_usage(self):
        receta = Receta.objects.create(
            nombre="Base QA Cadena Operativa Pendiente",
            hash_contenido="hash-chain-actions-001",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("8.500000"),
            rendimiento_unidad=self.unidad_kg,
        )
        insumo = Insumo.objects.create(
            nombre="Harina QA Cadena Operativa Pendiente",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unidad_kg,
            activo=True,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=insumo,
            insumo_texto=insumo.nombre,
            unidad=self.unidad_kg,
            unidad_texto="kg",
            cantidad=Decimal("2.500000"),
            costo_unitario_snapshot=Decimal("5.000000"),
            match_status=LineaReceta.STATUS_AUTO,
            match_method="MANUAL",
            match_score=100,
        )
        RecetaPresentacion.objects.create(
            receta=receta,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.380000"),
            activo=True,
        )

        response = self.client.get(reverse("recetas:receta_detail", args=[receta.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Controles de la cadena")
        self.assertContains(response, "Administrar presentaciones")
        self.assertContains(response, "Crear producto final")
        self.assertContains(response, "Aún no hay producto final consumiendo esta base.")


class RecetasAuthRedirectTests(TestCase):
    def test_drivers_legacy_redirects_to_login_when_anonymous(self):
        response = self.client.get(reverse("recetas:drivers_costeo_legacy"))
        self.assertEqual(response.status_code, 302)
        self.assertIn("/login/", response["Location"])
        self.assertIn("next=", response["Location"])

    def test_pronosticos_legacy_redirects_to_login_when_anonymous(self):
        response = self.client.get(reverse("recetas:pronosticos_legacy"))
        self.assertEqual(response.status_code, 302)
        self.assertIn("/login/", response["Location"])
        self.assertIn("next=", response["Location"])

    def test_matching_pendientes_redirects_to_login_when_anonymous(self):
        response = self.client.get(reverse("recetas:matching_pendientes"))
        self.assertEqual(response.status_code, 302)
        self.assertIn("/login/", response["Location"])
        self.assertIn("next=", response["Location"])

    def test_plan_produccion_redirects_to_login_when_anonymous(self):
        response = self.client.get(reverse("recetas:plan_produccion"))
        self.assertEqual(response.status_code, 302)
        self.assertIn("/login/", response["Location"])
        self.assertIn("next=", response["Location"])


class PlanProduccionRobustnessTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_plan_robust",
            email="admin_plan_robust@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

        self.unidad = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.insumo = Insumo.objects.create(nombre="Insumo robust", unidad_base=self.unidad, activo=True)
        self.receta = Receta.objects.create(nombre="Receta robust", hash_contenido="hash-robust-001")
        LineaReceta.objects.create(
            receta=self.receta,
            posicion=1,
            insumo=self.insumo,
            insumo_texto="Insumo robust",
            cantidad=Decimal("1"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("10"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        self.plan = PlanProduccion.objects.create(
            nombre="Plan robust",
            fecha_produccion=date(2026, 2, 20),
        )
        PlanProduccionItem.objects.create(
            plan=self.plan,
            receta=self.receta,
            cantidad=Decimal("2"),
        )

    def test_plan_produccion_graceful_when_pronostico_table_unavailable(self):
        with patch("recetas.views.plan.PronosticoVenta.objects.filter", side_effect=OperationalError("missing table")):
            response = self.client.get(reverse("recetas:plan_produccion"), {"plan_id": self.plan.id, "seccion": "todo"})

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["pronosticos_unavailable"])
        self.assertTrue(response.context["plan_vs_pronostico"]["pronosticos_unavailable"])


class PlanProduccionAdminComparativoTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_plan_admin",
            email="admin_plan_admin@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

        unidad = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        insumo = Insumo.objects.create(nombre="Insumo admin", unidad_base=unidad, activo=True)
        receta = Receta.objects.create(nombre="Receta admin", hash_contenido="hash-admin-001")
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=insumo,
            insumo_texto="Insumo admin",
            cantidad=Decimal("1"),
            unidad=unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        self.plan = PlanProduccion.objects.create(nombre="Plan admin", fecha_produccion=date(2026, 2, 10))
        PlanProduccionItem.objects.create(plan=self.plan, receta=receta, cantidad=Decimal("3"))
        PronosticoVenta.objects.create(receta=receta, periodo="2026-02", cantidad=Decimal("2"))

    def test_admin_comparativo_view_loads(self):
        response = self.client.get(
            reverse("admin:recetas_planproduccion_comparativo_pronostico", args=[self.plan.id])
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Comparativo Plan vs Pronóstico")
        self.assertContains(response, "Plan admin")
        self.assertContains(response, "Receta admin")

    def test_admin_changelist_shows_comparativo_link(self):
        response = self.client.get(reverse("admin:recetas_planproduccion_changelist"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ver plan vs pronóstico")


class PlanProduccionPeriodoMrpTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_plan_periodo",
            email="admin_plan_periodo@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

        unidad = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.insumo = Insumo.objects.create(nombre="Insumo periodo", unidad_base=unidad, activo=True)
        self.receta = Receta.objects.create(nombre="Receta periodo", hash_contenido="hash-periodo-001")
        LineaReceta.objects.create(
            receta=self.receta,
            posicion=1,
            insumo=self.insumo,
            insumo_texto="Insumo periodo",
            cantidad=Decimal("2"),
            unidad=unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("10"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        self.plan_q1 = PlanProduccion.objects.create(
            nombre="Plan q1",
            fecha_produccion=date(2026, 2, 10),
        )
        self.plan_q2 = PlanProduccion.objects.create(
            nombre="Plan q2",
            fecha_produccion=date(2026, 2, 22),
        )
        PlanProduccionItem.objects.create(plan=self.plan_q1, receta=self.receta, cantidad=Decimal("1"))
        PlanProduccionItem.objects.create(plan=self.plan_q2, receta=self.receta, cantidad=Decimal("3"))
        self.sucursal_a = Sucursal.objects.create(codigo="PLAN-A", nombre="Sucursal Plan A", activa=True)
        self.sucursal_b = Sucursal.objects.create(codigo="PLAN-B", nombre="Sucursal Plan B", activa=True)
        VentaHistorica.objects.create(
            receta=self.receta,
            sucursal=self.sucursal_a,
            fecha=date(2026, 1, 1),
            cantidad=Decimal("5"),
            monto_total=Decimal("500"),
            fuente="POINT_HIST_2026_Q1",
        )
        VentaHistorica.objects.create(
            receta=self.receta,
            sucursal=self.sucursal_b,
            fecha=date(2026, 1, 2),
            cantidad=Decimal("7"),
            monto_total=Decimal("700"),
            fuente="POINT_HIST_2026_Q1",
        )
        VentaHistorica.objects.create(
            receta=self.receta,
            sucursal=self.sucursal_a,
            fecha=date(2026, 1, 3),
            cantidad=Decimal("4"),
            monto_total=Decimal("400"),
            fuente="POINT_HIST_2026_Q1",
        )

    def test_plan_produccion_contexto_mrp_periodo_mes(self):
        response = self.client.get(
            reverse("recetas:plan_produccion"),
            {"mrp_periodo": "2026-02", "mrp_periodo_tipo": "mes"},
        )
        self.assertEqual(response.status_code, 200)
        resumen = response.context["mrp_periodo_resumen"]
        self.assertEqual(resumen["planes_count"], 2)
        self.assertEqual(resumen["insumos_count"], 1)
        self.assertEqual(resumen["periodo"], "2026-02")
        self.assertEqual(resumen["periodo_tipo"], "mes")
        self.assertEqual(resumen["insumos"][0]["cantidad"], Decimal("8"))
        self.assertEqual(resumen["costo_total"], Decimal("80"))
        self.assertEqual(resumen["health_label"], "Atención operativa")
        self.assertTrue(any(card["label"] == "Stock insuficiente" for card in resumen["quality_cards"]))
        self.assertTrue(any(card["label"] == "Maestro incompleto" for card in resumen["quality_cards"]))
        self.assertTrue(any(card["class_label"] == "Materia prima" for card in resumen["article_class_cards"]))
        self.assertTrue(any(row["scope"] == "Insumo" and row["label"] == "Stock insuficiente" for row in resumen["blocker_detail_rows"]))
        self.assertTrue(any(row["class_label"] == "Materia prima" for row in resumen["master_blocker_detail_rows"]))

    def test_plan_produccion_contexto_mrp_periodo_q1_filtra_planes(self):
        response = self.client.get(
            reverse("recetas:plan_produccion"),
            {"mrp_periodo": "2026-02", "mrp_periodo_tipo": "q1"},
        )
        self.assertEqual(response.status_code, 200)
        resumen = response.context["mrp_periodo_resumen"]
        self.assertEqual(resumen["planes_count"], 1)
        self.assertEqual(len(resumen["planes"]), 1)
        self.assertEqual(resumen["planes"][0]["id"], self.plan_q1.id)
        self.assertEqual(resumen["insumos"][0]["cantidad"], Decimal("2"))

    def test_plan_produccion_renders_mrp_periodo_enterprise_blockers(self):
        response = self.client.get(
            reverse("recetas:plan_produccion"),
            {"mrp_periodo": "2026-02", "mrp_periodo_tipo": "mes"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Bloqueos MRP del período")
        self.assertContains(response, "Abastecimiento por clase operativa")
        self.assertContains(response, "Detalle accionable MRP")
        self.assertContains(response, "Bloqueos del maestro")
        self.assertContains(response, "Bloqueos del maestro por dato faltante en MRP")
        self.assertContains(response, "Detalle del maestro bloqueando MRP")
        self.assertContains(response, "Stock insuficiente")
        self.assertContains(response, "Maestro incompleto")

    def test_plan_produccion_periodo_can_focus_quality_blocker(self):
        response = self.client.get(
            reverse("recetas:plan_produccion"),
            {
                "mrp_periodo": "2026-02",
                "mrp_periodo_tipo": "mes",
                "mrp_focus_kind": "quality",
                "mrp_focus_key": "stock_insuficiente",
            },
        )
        self.assertEqual(response.status_code, 200)
        resumen = response.context["mrp_periodo_resumen"]
        self.assertEqual(resumen["selected_focus_kind"], "quality")
        self.assertEqual(resumen["selected_focus_key"], "stock_insuficiente")
        self.assertIsNotNone(resumen["focus_summary"])
        self.assertTrue(resumen["insumos"])

    def test_plan_produccion_periodo_can_focus_master_missing_blocker(self):
        response = self.client.get(
            reverse("recetas:plan_produccion"),
            {
                "mrp_periodo": "2026-02",
                "mrp_periodo_tipo": "mes",
                "mrp_focus_kind": "master_missing",
                "mrp_focus_key": "proveedor",
            },
        )
        self.assertEqual(response.status_code, 200)
        resumen = response.context["mrp_periodo_resumen"]
        self.assertEqual(resumen["selected_focus_kind"], "master_missing")
        self.assertEqual(resumen["selected_focus_key"], "proveedor")
        self.assertTrue(resumen["master_blocker_detail_rows"])
        self.assertTrue(
            all("proveedor principal" in (row["missing"] or "").lower() for row in resumen["master_blocker_detail_rows"])
        )
        self.assertContains(response, "Bloqueos del maestro por dato faltante en MRP")
        self.assertContains(response, "kpi-card is-active", html=False)
        self.assertTrue(all(row["master_incomplete"] for row in resumen["insumos"]))
        self.assertContains(response, "Vista enfocada: proveedor principal")

    def test_plan_produccion_periodo_can_focus_master_class(self):
        response = self.client.get(
            reverse("recetas:plan_produccion"),
            {
                "mrp_periodo": "2026-02",
                "mrp_periodo_tipo": "mes",
                "mrp_focus_kind": "master",
                "mrp_focus_key": "MATERIA_PRIMA",
            },
        )
        self.assertEqual(response.status_code, 200)
        resumen = response.context["mrp_periodo_resumen"]
        self.assertEqual(resumen["selected_focus_kind"], "master")
        self.assertEqual(resumen["selected_focus_key"], "materia_prima")
        self.assertTrue(resumen["insumos"])
        self.assertTrue(all(row["master_incomplete"] for row in resumen["insumos"]))
        self.assertTrue(all(row["article_class_key"] == "MATERIA_PRIMA" for row in resumen["insumos"]))
        self.assertTrue(all(row["class_key"] == "MATERIA_PRIMA" for row in resumen["master_blocker_detail_rows"]))
        self.assertContains(response, "Vista enfocada:")

    def test_plan_produccion_periodo_export_csv(self):
        response = self.client.get(
            reverse("recetas:plan_produccion_periodo_export"),
            {"mrp_periodo": "2026-02", "mrp_periodo_tipo": "mes", "format": "csv"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        content = response.content.decode("utf-8")
        self.assertIn("MRP CONSOLIDADO POR PERIODO", content)
        self.assertIn("Insumo periodo", content)

    def test_plan_produccion_periodo_export_xlsx(self):
        response = self.client.get(
            reverse("recetas:plan_produccion_periodo_export"),
            {"mrp_periodo": "2026-02", "mrp_periodo_tipo": "q2", "format": "xlsx"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])
        wb = load_workbook(BytesIO(response.content), data_only=True)
        self.assertIn("Resumen", wb.sheetnames)
        self.assertIn("Planes", wb.sheetnames)
        self.assertIn("Insumos", wb.sheetnames)

    def test_plan_produccion_context_includes_enterprise_board(self):
        response = self.client.get(
            reverse("recetas:plan_produccion"),
            {
                "plan_id": self.plan_q1.id,
                "seccion": "todo",
                "mrp_periodo": "2026-02",
                "mrp_periodo_tipo": "mes",
            },
        )
        self.assertEqual(response.status_code, 200)
        board = response.context["enterprise_board"]
        self.assertIsNotNone(board)
        self.assertEqual(board["release_label"], "Listo para compras")
        self.assertEqual(board["ready_for_purchase_total"], 1)
        blocker_labels = [item["label"] for item in board["blocker_cards"]]
        self.assertIn("Stock insuficiente", blocker_labels)
        self.assertIn("Desviaciones forecast", blocker_labels)
        detail_rows = board["blocker_detail_rows"]
        self.assertTrue(any(row["scope"] == "Insumo" and row["label"] == "Sin proveedor" for row in detail_rows))
        self.assertTrue(any(row["scope"] == "Forecast" for row in detail_rows))
        master_cards = board["master_blocker_class_cards"]
        self.assertTrue(any(card["class_label"] == "Materia prima" for card in master_cards))
        master_rows = board["master_blocker_detail_rows"]
        self.assertTrue(any(row["class_label"] == "Materia prima" and "proveedor principal" in row["missing"] for row in master_rows))
        supply_rows = response.context["explosion"]["insumos"]
        self.assertEqual(supply_rows[0]["workflow_health_label"], "Sin proveedor")

    def test_plan_produccion_renders_enterprise_sections(self):
        response = self.client.get(
            reverse("recetas:plan_produccion"),
            {
                "plan_id": self.plan_q1.id,
                "seccion": "diagnostico",
                "mrp_periodo": "2026-02",
                "mrp_periodo_tipo": "mes",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Control del plan")
        self.assertContains(response, "Bloqueos ERP")
        self.assertContains(response, "Siguiente paso")
        self.assertContains(response, "Paso de Compras")
        self.assertContains(response, "Salud ERP")
        self.assertContains(response, "Lista para operar")
        self.assertContains(response, "Sin proveedor")
        self.assertContains(response, "Detalle accionable de bloqueos")
        self.assertContains(response, "Elemento bloqueado")
        self.assertContains(response, "Asignar proveedor")
        self.assertContains(response, "Bloqueos del maestro por clase")
        self.assertContains(response, "Detalle del maestro bloqueando el plan")
        self.assertContains(response, "Faltante maestro")
        self.assertContains(response, "Cargar pronóstico")

    def test_plan_produccion_renders_sales_history_summary(self):
        cache.clear()
        response = self.client.get(
            reverse("recetas:plan_produccion"),
            {
                "plan_id": self.plan_q1.id,
                "seccion": "todo",
                "mrp_periodo": "2026-02",
                "mrp_periodo_tipo": "mes",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cobertura de demanda")
        self.assertContains(response, "Sucursales líderes")
        self.assertContains(response, "Productos líderes")
        summary = response.context["ventas_historicas_summary"]
        self.assertTrue(summary["available"])
        self.assertEqual(summary["branch_count"], 2)
        self.assertEqual(summary["recipe_count"], 1)
        self.assertEqual(summary["active_days"], 3)
        self.assertEqual(summary["expected_days"], 3)
        self.assertEqual(summary["missing_days"], 0)
        self.assertEqual(summary["top_branches"][0]["sucursal__codigo"], "PLAN-A")
        self.assertContains(response, "Top decisiones del plan")
        self.assertContains(response, "Sucursales que empujan el plan")
        self.assertContains(response, "Insumo a asegurar por sucursal")
        self.assertContains(response, "PLAN-A")
        self.assertContains(response, self.receta.nombre)
        self.assertIn("daily_decision_rows", response.context)
        self.assertTrue(response.context["daily_decision_rows"])
        self.assertIn("branch_priority_rows", response.context)
        self.assertTrue(response.context["branch_priority_rows"])
        self.assertIn("branch_supply_rows", response.context)
        self.assertTrue(response.context["branch_supply_rows"])
        self.assertEqual(response.context["branch_priority_rows"][0]["dominant_recipe_name"], self.receta.nombre)


class PlanProduccionWorkflowRowsTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_plan_workflow",
            email="admin_plan_workflow@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

        self.unidad = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.insumo = Insumo.objects.create(nombre="Insumo WF", unidad_base=self.unidad, activo=True)
        self.receta_ok = Receta.objects.create(nombre="Receta WF OK", hash_contenido="hash-wf-ok")
        self.receta_bad = Receta.objects.create(nombre="Receta WF Bad", hash_contenido="hash-wf-bad")
        LineaReceta.objects.create(
            receta=self.receta_ok,
            posicion=1,
            insumo=self.insumo,
            insumo_texto="Insumo WF",
            cantidad=Decimal("1"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        LineaReceta.objects.create(
            receta=self.receta_bad,
            posicion=1,
            insumo=None,
            insumo_texto="Insumo sin match",
            cantidad=Decimal("1"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("0"),
            match_status=LineaReceta.STATUS_REJECTED,
            match_score=0,
            match_method=LineaReceta.MATCH_NONE,
        )
        self.plan = PlanProduccion.objects.create(nombre="Plan WF", fecha_produccion=date(2026, 2, 12))
        PlanProduccionItem.objects.create(plan=self.plan, receta=self.receta_ok, cantidad=Decimal("1"))
        PlanProduccionItem.objects.create(plan=self.plan, receta=self.receta_bad, cantidad=Decimal("1"))

    def test_plan_produccion_item_rows_include_workflow_health(self):
        response = self.client.get(reverse("recetas:plan_produccion"), {"plan_id": self.plan.id, "seccion": "diagnostico"})
        self.assertEqual(response.status_code, 200)
        items = response.context["explosion"]["items_detalle"]
        labels = {row["receta"].nombre: row["workflow_health_label"] for row in items}
        self.assertEqual(labels["Receta WF OK"], "Lista para operar")
        self.assertEqual(labels["Receta WF Bad"], "Sin artículo estándar")
        self.assertContains(response, "Resolver catálogo")


class PlanProduccionSolicitudesModeTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_plan",
            email="admin_plan@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

        self.unidad = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.proveedor = Proveedor.objects.create(nombre="Proveedor Plan", activo=True)
        self.sucursal = Sucursal.objects.create(codigo="PLAN", nombre="Sucursal Plan", activa=True)
        self.insumo = Insumo.objects.create(
            nombre="Harina Plan",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        self.receta = Receta.objects.create(nombre="Receta Plan", hash_contenido="hash-plan-001")
        LineaReceta.objects.create(
            receta=self.receta,
            posicion=1,
            insumo=self.insumo,
            insumo_texto="Harina Plan",
            cantidad=Decimal("2"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        self.plan = PlanProduccion.objects.create(nombre="Plan Test", fecha_produccion=date(2026, 2, 20))
        PlanProduccionItem.objects.create(plan=self.plan, receta=self.receta, cantidad=Decimal("1"))

    def test_generar_solicitudes_accumulate_mode_updates_existing(self):
        url = reverse("recetas:plan_produccion_generar_solicitudes", args=[self.plan.id])

        response_1 = self.client.post(
            url,
            {"next_view": "plan", "replace_prev": "1", "auto_create_oc": "1"},
        )
        self.assertEqual(response_1.status_code, 302)

        response_2 = self.client.post(
            url,
            {"next_view": "plan", "replace_prev": "0", "auto_create_oc": "1"},
        )
        self.assertEqual(response_2.status_code, 302)

        solicitudes = SolicitudCompra.objects.filter(
            area=f"PLAN_PRODUCCION:{self.plan.id}",
            estatus=SolicitudCompra.STATUS_BORRADOR,
            insumo=self.insumo,
        )
        self.assertEqual(solicitudes.count(), 1)
        self.assertEqual(solicitudes.first().cantidad, Decimal("4"))

        ocs = OrdenCompra.objects.filter(
            referencia=f"PLAN_PRODUCCION:{self.plan.id}",
            estatus=OrdenCompra.STATUS_BORRADOR,
            solicitud__isnull=True,
            proveedor=self.proveedor,
        )
        self.assertEqual(ocs.count(), 1)
        self.assertEqual(ocs.first().monto_estimado, Decimal("20"))

    def test_plan_produccion_context_includes_document_control(self):
        insumo_blocked = Insumo.objects.create(
            nombre="Chocolate plan sin proveedor",
            categoria="Cobertura",
            unidad_base=self.unidad,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=insumo_blocked,
            proveedor=self.proveedor,
            costo_unitario=Decimal("12.50"),
            source_hash="cost-plan-doc-master-blocker",
        )
        SolicitudCompra.objects.create(
            area=f"PLAN_PRODUCCION:{self.plan.id}",
            solicitante="Plan bloqueado",
            insumo=insumo_blocked,
            cantidad=Decimal("1.000"),
            fecha_requerida=self.plan.fecha_produccion,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        solicitud = SolicitudCompra.objects.create(
            area=f"PLAN_PRODUCCION:{self.plan.id}",
            solicitante="Plan automático",
            insumo=self.insumo,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("2.000"),
            fecha_requerida=self.plan.fecha_produccion,
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        orden = OrdenCompra.objects.create(
            solicitud=solicitud,
            referencia=f"PLAN_PRODUCCION:{self.plan.id}",
            proveedor=self.proveedor,
            fecha_entrega_estimada=self.plan.fecha_produccion,
            monto_estimado=Decimal("10.00"),
            estatus=OrdenCompra.STATUS_CONFIRMADA,
        )
        RecepcionCompra.objects.create(
            orden=orden,
            fecha_recepcion=self.plan.fecha_produccion,
            estatus=RecepcionCompra.STATUS_PENDIENTE,
            conformidad_pct=Decimal("100.00"),
        )

        response = self.client.get(reverse("recetas:plan_produccion"), {"plan_id": self.plan.id, "seccion": "diagnostico"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Prioridades de atención")
        self.assertContains(response, "Resumen de seguimiento")
        document_control = response.context["document_control"]
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertIsNotNone(document_control)
        self.assertEqual(document_control["solicitudes_total"], 2)
        self.assertEqual(document_control["ordenes_total"], 1)
        self.assertEqual(document_control["recepciones_total"], 1)
        self.assertEqual(document_control["solicitudes_pendientes_total"], 1)
        self.assertEqual(document_control["ordenes_pendientes_total"], 0)
        self.assertEqual(document_control["recepciones_abiertas_total"], 1)
        self.assertEqual(document_control["blocked_total"], 2)
        self.assertEqual(document_control["health_label"], "Con bloqueos")
        self.assertEqual(document_control["stage_label"], "Recepción en proceso")
        self.assertEqual(document_control["next_action_label"], "Abrir recepciones")
        health_cards = {card["label"]: card for card in document_control["document_health_cards"]}
        stage_rows = {row["label"]: row for row in document_control["document_stage_rows"]}
        pipeline_steps = {step["label"]: step for step in document_control["pipeline_steps"]}
        blocker_rows = document_control["document_blocker_rows"]
        self.assertEqual(health_cards["Recepciones abiertas"]["count"], 1)
        self.assertEqual(health_cards["Solicitudes pendientes"]["count"], 1)
        self.assertEqual(stage_rows["Solicitudes"]["closed_count"], 1)
        self.assertEqual(stage_rows["Solicitudes"]["open_count"], 1)
        self.assertEqual(stage_rows["Solicitudes"]["semaphore_label"], "Amarillo")
        self.assertEqual(stage_rows["Solicitudes"]["progress_pct"], 50)
        self.assertEqual(stage_rows["Recepciones"]["open_count"], 1)
        self.assertEqual(stage_rows["Recepciones"]["semaphore_label"], "Rojo")
        self.assertEqual(stage_rows["Recepciones"]["progress_pct"], 0)
        self.assertEqual(pipeline_steps["Solicitudes"]["status_label"], "Por atender")
        self.assertEqual(pipeline_steps["Solicitudes"]["semaphore_label"], "Amarillo")
        self.assertEqual(pipeline_steps["Solicitudes"]["action_label"], "Liberar solicitudes")
        self.assertIn("Aprueba o termina la captura", pipeline_steps["Solicitudes"]["action_detail"])
        self.assertEqual(pipeline_steps["Recepciones"]["status_label"], "Abiertas")
        self.assertEqual(pipeline_steps["Recepciones"]["semaphore_label"], "Rojo")
        self.assertEqual(pipeline_steps["Recepciones"]["blocked"], 2)
        self.assertGreaterEqual(pipeline_steps["Recepciones"]["progress_pct"], 0)
        self.assertEqual(pipeline_steps["Recepciones"]["action_label"], "Cerrar recepciones")
        self.assertIn("Aplica inventario", pipeline_steps["Recepciones"]["action_detail"])
        self.assertEqual(document_control["closure_checks"][2]["label"], "Recepciones aplicadas")
        self.assertEqual(document_control["closure_checks"][2]["action_label"], "Cerrar recepciones abiertas")
        self.assertIn("Aplica inventario", document_control["closure_checks"][2]["action_detail"])
        self.assertIn("closure_key=recepciones_aplicadas", document_control["closure_checks"][2]["focus_url"])
        self.assertEqual(document_control["closure_summary"]["label"], "Cierre documental pendiente")
        self.assertEqual(document_control["closure_summary"]["ready_count"], 1)
        self.assertEqual(document_control["closure_summary"]["pending_count"], 2)
        self.assertEqual(document_control["closure_summary"]["progress_pct"], 33)
        self.assertEqual(document_control["closure_focus"]["label"], "Recepciones aplicadas")
        self.assertEqual(document_control["closure_focus"]["tone"], "danger")
        self.assertTrue(document_control["closure_focus_rows"])
        self.assertEqual(document_control["closure_focus_rows"][0]["scope"], "Recepción")
        self.assertEqual([item["label"] for item in document_control["handoff_checks"]], ["Solicitud → Orden", "Orden → Recepción", "Recepción → Cierre"])
        self.assertEqual(document_control["handoff_checks"][2]["action_label"], "Cerrar recepciones")
        self.assertIn("handoff_key=recepcion_cierre", document_control["handoff_checks"][2]["focus_url"])
        self.assertEqual(document_control["handoff_summary"]["label"], "Entregas entre etapas pendientes")
        self.assertEqual(document_control["handoff_summary"]["ready_count"], 0)
        self.assertEqual(document_control["handoff_summary"]["pending_count"], 3)
        self.assertEqual(document_control["handoff_summary"]["blocked_count"], 2)
        self.assertEqual(document_control["handoff_summary"]["progress_pct"], 0)
        self.assertEqual(document_control["handoff_focus"]["label"], "Recepción → Cierre")
        self.assertEqual(document_control["handoff_focus"]["tone"], "danger")
        self.assertTrue(document_control["handoff_focus_rows"])
        self.assertEqual(document_control["handoff_focus_rows"][0]["scope"], "Recepción")
        self.assertEqual(document_control["selected_stage_key"], "auto")
        self.assertEqual(document_control["selected_closure_key"], "auto")
        self.assertEqual(document_control["selected_handoff_key"], "auto")
        self.assertEqual(document_control["stage_focus"]["scope"], "Recepción")
        self.assertTrue(document_control["stage_focus"]["blocker_rows"])
        self.assertTrue(all(row["scope"] == "Recepción" for row in document_control["stage_focus"]["blocker_rows"]))
        self.assertTrue(any(row["scope"] == "Recepción" and row["folio"] for row in blocker_rows))
        self.assertEqual(document_control["purchase_gate"]["label"], "Bloqueado")
        self.assertIn("bloqueos documentales", document_control["purchase_gate"]["detail"].lower())
        self.assertIn(f"source=plan&plan_id={self.plan.id}", document_control["solicitudes_url"])
        self.assertIn(f"source=plan&plan_id={self.plan.id}", document_control["ordenes_url"])
        self.assertIn(f"source=plan&plan_id={self.plan.id}", document_control["recepciones_url"])
        self.assertTrue(document_control["master_blocker_class_cards"])
        self.assertTrue(document_control["master_blocker_missing_cards"])
        self.assertTrue(document_control["master_blocker_detail_rows"])
        self.assertGreaterEqual(document_control["master_blocker_total"], 1)
        self.assertEqual(document_control["master_focus"]["class_label"], "Materia prima")
        self.assertIn(document_control["master_focus"]["missing_field"], {"proveedor principal", "código Point", "código comercial", "código externo"})
        self.assertTrue(document_control["master_focus_rows"])
        self.assertTrue(
            any((card.get("action_label") or "").strip() for card in document_control["master_blocker_class_cards"])
        )
        self.assertTrue(
            any((card.get("action_detail") or "").strip() for card in document_control["master_blocker_class_cards"])
        )
        self.assertTrue(
            any((row.get("action_label") or "").strip() for row in document_control["master_blocker_detail_rows"])
        )
        self.assertTrue(
            any((row.get("action_detail") or "").strip() for row in document_control["master_blocker_detail_rows"])
        )
        self.assertTrue(
            any((card.get("missing_label") or "").strip() for card in document_control["master_blocker_missing_cards"])
        )
        self.assertTrue(
            any((card.get("action_label") or "").strip() for card in document_control["master_blocker_missing_cards"])
        )

    def test_plan_produccion_can_focus_stage_closure_and_handoff(self):
        solicitud = SolicitudCompra.objects.create(
            area=f"PLAN_PRODUCCION:{self.plan.id}",
            solicitante="Plan enfoque",
            insumo=self.insumo,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("2.000"),
            fecha_requerida=self.plan.fecha_produccion,
            estatus=SolicitudCompra.STATUS_APROBADA,
        )
        orden = OrdenCompra.objects.create(
            solicitud=solicitud,
            referencia=f"PLAN_PRODUCCION:{self.plan.id}",
            proveedor=self.proveedor,
            fecha_entrega_estimada=self.plan.fecha_produccion,
            monto_estimado=Decimal("10.00"),
            estatus=OrdenCompra.STATUS_CONFIRMADA,
        )
        RecepcionCompra.objects.create(
            orden=orden,
            fecha_recepcion=self.plan.fecha_produccion,
            estatus=RecepcionCompra.STATUS_PENDIENTE,
            conformidad_pct=Decimal("100.00"),
        )

        response = self.client.get(
            reverse("recetas:plan_produccion"),
            {
                "plan_id": self.plan.id,
                "stage_key": "ordenes",
                "closure_key": "ordenes_sin_bloqueo",
                "handoff_key": "orden_recepcion",
            },
        )
        self.assertEqual(response.status_code, 200)
        document_control = response.context["document_control"]
        self.assertEqual(document_control["selected_stage_key"], "ordenes")
        self.assertEqual(document_control["selected_closure_key"], "ordenes_sin_bloqueo")
        self.assertEqual(document_control["selected_handoff_key"], "orden_recepcion")
        self.assertEqual(document_control["stage_focus"]["key"], "ordenes")
        self.assertEqual(document_control["closure_focus"]["key"], "ordenes_sin_bloqueo")
        self.assertEqual(document_control["handoff_focus"]["key"], "orden_recepcion")
        self.assertContains(response, "Enfocar")
        self.assertContains(response, "kpi-card is-active", html=False)

    def test_plan_produccion_can_focus_master_blocker_class(self):
        insumo_blocked = Insumo.objects.create(
            nombre="Etiqueta plan foco master",
            categoria="",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor,
            activo=True,
            tipo_item=Insumo.TIPO_EMPAQUE,
        )
        CostoInsumo.objects.create(
            insumo=insumo_blocked,
            proveedor=self.proveedor,
            costo_unitario=Decimal("2.20"),
            source_hash="cost-plan-doc-master-focus",
        )
        SolicitudCompra.objects.create(
            area=f"PLAN_PRODUCCION:{self.plan.id}",
            solicitante="Plan bloqueado focus",
            insumo=insumo_blocked,
            cantidad=Decimal("1.000"),
            fecha_requerida=self.plan.fecha_produccion,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )

        response = self.client.get(
            reverse("recetas:plan_produccion"),
            {
                "plan_id": self.plan.id,
                "master_focus_key": "EMPAQUE",
            },
        )
        self.assertEqual(response.status_code, 200)
        document_control = response.context["document_control"]
        self.assertEqual(document_control["selected_master_focus_key"], "empaque")
        self.assertTrue(document_control["master_focus_rows"])
        self.assertTrue(all(row["class_label"] == "Empaque" for row in document_control["master_focus_rows"]))
        focus_row = document_control["master_focus_rows"][0]
        self.assertIn(f"insumo_id={insumo_blocked.id}", focus_row["action_url"])
        self.assertEqual(focus_row["edit_url"], reverse("maestros:insumo_update", args=[insumo_blocked.id]))
        self.assertEqual(
            document_control["master_focus"]["edit_url"],
            reverse("maestros:insumo_update", args=[insumo_blocked.id]),
        )
        self.assertContains(response, "Vista enfocada")
        self.assertContains(response, "Editar artículo")
        self.assertContains(response, "kpi-card is-active", html=False)

    def test_plan_produccion_can_focus_master_blocker_missing_field(self):
        insumo_blocked = Insumo.objects.create(
            nombre="Etiqueta plan foco missing",
            categoria="",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor,
            activo=True,
            tipo_item=Insumo.TIPO_EMPAQUE,
        )
        CostoInsumo.objects.create(
            insumo=insumo_blocked,
            proveedor=self.proveedor,
            costo_unitario=Decimal("2.20"),
            source_hash="cost-plan-doc-master-missing-focus",
        )
        SolicitudCompra.objects.create(
            area=f"PLAN_PRODUCCION:{self.plan.id}",
            solicitante="Plan bloqueado focus missing",
            insumo=insumo_blocked,
            cantidad=Decimal("1.000"),
            fecha_requerida=self.plan.fecha_produccion,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )

        response = self.client.get(
            reverse("recetas:plan_produccion"),
            {
                "plan_id": self.plan.id,
                "master_missing_key": "categoria",
            },
        )
        self.assertEqual(response.status_code, 200)
        document_control = response.context["document_control"]
        self.assertEqual(document_control["selected_master_missing_key"], "categoria")
        self.assertTrue(document_control["master_focus_rows"])
        self.assertTrue(
            all("categoría" in (row["missing_field"] or "").lower() for row in document_control["master_focus_rows"])
        )
        focus_row = document_control["master_focus_rows"][0]
        self.assertIn(f"insumo_id={insumo_blocked.id}", focus_row["action_url"])
        self.assertEqual(focus_row["edit_url"], reverse("maestros:insumo_update", args=[insumo_blocked.id]))
        self.assertEqual(
            document_control["master_focus"]["edit_url"],
            reverse("maestros:insumo_update", args=[insumo_blocked.id]),
        )
        self.assertContains(response, "Bloqueos del maestro por dato faltante")
        self.assertContains(response, "Editar artículo")
        self.assertContains(response, "kpi-card is-active", html=False)

    def test_plan_produccion_renders_document_control(self):
        insumo_blocked = Insumo.objects.create(
            nombre="Etiqueta plan sin categoria",
            categoria="",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor,
            activo=True,
            tipo_item=Insumo.TIPO_EMPAQUE,
        )
        CostoInsumo.objects.create(
            insumo=insumo_blocked,
            proveedor=self.proveedor,
            costo_unitario=Decimal("2.20"),
            source_hash="cost-plan-doc-render-blocker",
        )
        SolicitudCompra.objects.create(
            area=f"PLAN_PRODUCCION:{self.plan.id}",
            solicitante="Plan bloqueado render",
            insumo=insumo_blocked,
            cantidad=Decimal("1.000"),
            fecha_requerida=self.plan.fecha_produccion,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        solicitud = SolicitudCompra.objects.create(
            area=f"PLAN_PRODUCCION:{self.plan.id}",
            solicitante="Plan automático",
            insumo=self.insumo,
            proveedor_sugerido=self.proveedor,
            cantidad=Decimal("2.000"),
            fecha_requerida=self.plan.fecha_produccion,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        orden = OrdenCompra.objects.create(
            solicitud=solicitud,
            referencia=f"PLAN_PRODUCCION:{self.plan.id}",
            proveedor=self.proveedor,
            fecha_entrega_estimada=self.plan.fecha_produccion,
            monto_estimado=Decimal("10.00"),
            estatus=OrdenCompra.STATUS_CONFIRMADA,
        )
        RecepcionCompra.objects.create(
            orden=orden,
            fecha_recepcion=self.plan.fecha_produccion,
            estatus=RecepcionCompra.STATUS_CERRADA,
            conformidad_pct=Decimal("100.00"),
        )

        response = self.client.get(reverse("recetas:plan_produccion"), {"plan_id": self.plan.id, "seccion": "diagnostico"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Control Documental")
        self.assertContains(response, "Plan / BOM")
        self.assertContains(response, "Compras documentales")
        self.assertContains(response, "Inventario / Reabasto")
        self.assertContains(response, "Control de demanda comercial")
        self.assertContains(response, "Control de maestro crítico por demanda")
        self.assertIn("demand_gate_summary", response.context)
        self.assertIn("master_demand_gate_summary", response.context)
        self.assertContains(response, "Control por frente")
        self.assertContains(response, "Solicitudes")
        self.assertContains(response, "Órdenes")
        self.assertContains(response, "Recepciones")
        self.assertContains(response, "Solicitudes pendientes")
        self.assertContains(response, "Órdenes por confirmar")
        self.assertContains(response, "Recepciones abiertas")
        self.assertContains(response, "Documentos cerrados")
        self.assertContains(response, "Cierre documental pendiente")
        self.assertContains(response, "Entregas entre etapas pendientes")
        self.assertContains(response, "bloqueados")
        self.assertContains(response, "Cierre")
        self.assertContains(response, "Cerradas")
        self.assertContains(response, "Etapa prioritaria")
        self.assertContains(response, "Bloqueos del maestro por dato faltante")
        self.assertContains(response, "Folio")
        self.assertContains(response, "Estatus")
        self.assertContains(response, "Documento")
        self.assertContains(response, "Semáforo")
        self.assertContains(response, "Avance")
        self.assertContains(response, "Criterio de cierre prioritario")
        self.assertContains(response, "Entregas entre etapas")
        self.assertContains(response, "Entrega prioritaria")

    def test_plan_produccion_flags_critical_master_demand_gate(self):
        insumo_critico = Insumo.objects.create(
            nombre="Caja critica plan",
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        receta_critica = Receta.objects.create(
            nombre="Pastel critico plan",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Pastel",
            hash_contenido="hash-plan-critico-master-001",
        )
        LineaReceta.objects.create(
            receta=receta_critica,
            posicion=1,
            insumo=insumo_critico,
            insumo_texto=insumo_critico.nombre,
            cantidad=Decimal("1"),
            unidad=self.unidad,
            unidad_texto="kg",
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        PlanProduccionItem.objects.create(plan=self.plan, receta=receta_critica, cantidad=Decimal("12"))
        VentaHistorica.objects.create(
            receta=receta_critica,
            sucursal=self.sucursal,
            fecha=self.plan.fecha_produccion - timedelta(days=2),
            cantidad=Decimal("94"),
            tickets=5,
        )

        response = self.client.get(reverse("recetas:plan_produccion"), {"plan_id": self.plan.id, "seccion": "diagnostico"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Control de maestro crítico por demanda")
        self.assertContains(response, "Liberación del plan retenida")
        self.assertContains(response, "Caja critica plan")
        self.assertIn("critical_master_demand_rows", response.context)
        self.assertTrue(response.context["critical_master_demand_rows"])
        plan_row = next(row for row in response.context["trunk_handoff_rows"] if row["label"] == "Plan / BOM")
        self.assertEqual(plan_row["tone"], "danger")
        self.assertEqual(plan_row["status"], "Crítico")
        self.assertContains(response, "Control de maestro crítico por demanda")
        self.assertContains(response, "Recepciones aplicadas")
        self.assertContains(response, "Recepción → Cierre")
        self.assertContains(response, "Recepciones al día")
        self.assertContains(response, "Las recepciones ya quedaron cerradas y aplicadas.")
        self.assertIn("trunk_handoff_rows", response.context)
        self.assertEqual(len(response.context["trunk_handoff_rows"]), 3)
        self.assertContains(response, "Cerrado")
        self.assertContains(response, f"source=plan&amp;plan_id={self.plan.id}")
        self.assertContains(response, f"source=plan&amp;plan_id={self.plan.id}")


class RecetaCosteoVersionadoTests(TestCase):
    def setUp(self):
        self.unidad = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.insumo = Insumo.objects.create(nombre="Harina Base", unidad_base=self.unidad, activo=True)
        self.receta = Receta.objects.create(
            nombre="Batida Vainilla",
            sheet_name="Insumos 1",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("10"),
            rendimiento_unidad=self.unidad,
            hash_contenido="hash-versionado-001",
        )
        self.linea = LineaReceta.objects.create(
            receta=self.receta,
            posicion=1,
            insumo=self.insumo,
            insumo_texto="Harina Base",
            cantidad=Decimal("2"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("4"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

    def test_aplica_driver_producto_y_versiona_por_cambio(self):
        CostoDriver.objects.create(
            nombre="Driver Producto Batida",
            scope=CostoDriver.SCOPE_PRODUCTO,
            receta=self.receta,
            mo_pct=Decimal("10"),
            indirecto_pct=Decimal("5"),
            mo_fijo=Decimal("1"),
            indirecto_fijo=Decimal("2"),
            prioridad=10,
        )

        v1, created_1 = asegurar_version_costeo(self.receta, fuente="TEST")
        self.assertTrue(created_1)
        self.assertEqual(v1.version_num, 1)
        self.assertEqual(v1.costo_mp, Decimal("8.000000"))
        self.assertEqual(v1.costo_mo, Decimal("1.800000"))
        self.assertEqual(v1.costo_indirecto, Decimal("2.400000"))
        self.assertEqual(v1.costo_total, Decimal("12.200000"))
        self.assertEqual(v1.costo_por_unidad_rendimiento, Decimal("1.220000"))

        self.linea.costo_unitario_snapshot = Decimal("5")
        self.linea.save(update_fields=["costo_unitario_snapshot"])
        v2, created_2 = asegurar_version_costeo(self.receta, fuente="TEST")
        self.assertTrue(created_2)
        self.assertEqual(v2.version_num, 2)
        self.assertEqual(v2.costo_mp, Decimal("10.000000"))
        self.assertEqual(v2.costo_total, Decimal("14.500000"))

        self.assertEqual(RecetaCostoVersion.objects.filter(receta=self.receta).count(), 2)

    def test_calculo_no_driver_solo_mp(self):
        breakdown = calcular_costeo_receta(self.receta)
        self.assertEqual(breakdown.costo_mp, Decimal("8.000000"))
        self.assertEqual(breakdown.costo_mo, Decimal("0.000000"))
        self.assertEqual(breakdown.costo_indirecto, Decimal("0.000000"))
        self.assertEqual(breakdown.costo_total, Decimal("8.000000"))


class CosteoSnapshotResolverTests(TestCase):
    def setUp(self):
        self.unidad_g = UnidadMedida.objects.create(
            codigo="g",
            nombre="Gramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1"),
        )
        self.unidad_kg = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.unidad_ml = UnidadMedida.objects.create(
            codigo="ml",
            nombre="Mililitro",
            tipo=UnidadMedida.TIPO_VOLUMEN,
            factor_to_base=Decimal("1"),
        )
        self.unidad_lt = UnidadMedida.objects.create(
            codigo="lt",
            nombre="Litro",
            tipo=UnidadMedida.TIPO_VOLUMEN,
            factor_to_base=Decimal("1000"),
        )
        self.unidad_pza = UnidadMedida.objects.create(
            codigo="pza",
            nombre="Pieza",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        self.unidad_u = UnidadMedida.objects.create(
            codigo="unidad",
            nombre="Unidad",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )

    def test_resolve_line_snapshot_cost_converts_from_recipe_yield_unit(self):
        insumo = Insumo.objects.create(
            nombre="Mezcla 3 Leches",
            codigo_point="01M3L04",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_lt,
            activo=True,
        )
        receta = Receta.objects.create(
            nombre="Mezcla 3 Leches",
            codigo_point="01M3L04",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("1"),
            rendimiento_unidad=self.unidad_lt,
            hash_contenido="hash-cost-snapshot-lt-001",
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo_texto="Leche",
            insumo=Insumo.objects.create(nombre="Leche Base", unidad_base=self.unidad_lt, activo=True),
            cantidad=Decimal("1"),
            unidad=self.unidad_lt,
            unidad_texto="lt",
            costo_unitario_snapshot=Decimal("790.294004"),
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100,
        )
        linea = LineaReceta.objects.create(
            receta=Receta.objects.create(
                nombre="Pastel Test",
                codigo_point="TEST-ML",
                tipo=Receta.TIPO_PRODUCTO_FINAL,
                hash_contenido="hash-cost-snapshot-line-ml-001",
            ),
            posicion=1,
            insumo=insumo,
            insumo_texto=insumo.nombre,
            cantidad=Decimal("200"),
            unidad=self.unidad_ml,
            unidad_texto="ml",
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100,
        )

        snapshot, source = resolve_line_snapshot_cost(linea)
        self.assertEqual(source, "RECETA_PREPARACION")
        self.assertEqual(snapshot, Decimal("0.790294"))

    def test_resolve_line_snapshot_cost_converts_from_preparation_piece_unit(self):
        insumo = Insumo.objects.create(
            nombre="Pan Vainilla Dawn Mediano",
            codigo_point="01VDM",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_pza,
            activo=True,
        )
        receta = Receta.objects.create(
            nombre="Pan Vainilla Dawn Mediano",
            codigo_point="01VDM",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("1"),
            rendimiento_unidad=self.unidad_u,
            hash_contenido="hash-cost-snapshot-unit-001",
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo_texto="Harina",
            insumo=Insumo.objects.create(nombre="Harina Pan", unidad_base=self.unidad_kg, activo=True),
            cantidad=Decimal("1"),
            unidad=self.unidad_u,
            unidad_texto="unidad",
            costo_unitario_snapshot=Decimal("20.645793"),
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100,
        )
        linea = LineaReceta.objects.create(
            receta=Receta.objects.create(
                nombre="Pastel Test Pza",
                codigo_point="TEST-U",
                tipo=Receta.TIPO_PRODUCTO_FINAL,
                hash_contenido="hash-cost-snapshot-line-u-001",
            ),
            posicion=1,
            insumo=insumo,
            insumo_texto=insumo.nombre,
            cantidad=Decimal("2"),
            unidad=self.unidad_u,
            unidad_texto="U",
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100,
        )

        snapshot, source = resolve_line_snapshot_cost(linea)
        self.assertEqual(source, "RECETA_PREPARACION")
        self.assertEqual(snapshot, Decimal("20.645793"))


class PronosticoImportViewTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_pronostico",
            email="admin_pronostico@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)
        self.receta = Receta.objects.create(
            nombre="Pastel Fresas Con Crema - Chico",
            codigo_point="PFC-CHICO",
            hash_contenido="hash-pronostico-001",
        )

    def test_import_replace_crea_pronosticos(self):
        csv_data = (
            "receta,codigo_point,periodo,cantidad\n"
            "Pastel Fresas Con Crema - Chico,PFC-CHICO,2026-02,120\n"
        ).encode("utf-8")
        upload = SimpleUploadedFile("pronostico.csv", csv_data, content_type="text/csv")
        response = self.client.post(
            reverse("recetas:pronosticos_importar"),
            {
                "modo": "replace",
                "periodo_default": "2026-02",
                "fuente": "TEST_UI",
                "archivo": upload,
            },
        )
        self.assertEqual(response.status_code, 302)
        record = PronosticoVenta.objects.get(receta=self.receta, periodo="2026-02")
        self.assertEqual(record.cantidad, Decimal("120"))
        self.assertEqual(record.fuente, "TEST_UI")

    def test_import_accumulate_acumula_cantidad(self):
        PronosticoVenta.objects.create(
            receta=self.receta,
            periodo="2026-02",
            cantidad=Decimal("100"),
            fuente="MANUAL",
        )
        csv_data = (
            "codigo_point,periodo,cantidad\n"
            "PFC-CHICO,2026-02,20\n"
        ).encode("utf-8")
        upload = SimpleUploadedFile("pronostico.csv", csv_data, content_type="text/csv")
        response = self.client.post(
            reverse("recetas:pronosticos_importar"),
            {
                "modo": "accumulate",
                "periodo_default": "2026-02",
                "fuente": "TEST_UI",
                "archivo": upload,
            },
        )
        self.assertEqual(response.status_code, 302)
        record = PronosticoVenta.objects.get(receta=self.receta, periodo="2026-02")
        self.assertEqual(record.cantidad, Decimal("120"))


class PlanGeneradoDesdePronosticoTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_plan_gen",
            email="admin_plan_gen@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

        self.receta_final = Receta.objects.create(
            nombre="Producto Final Test",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-plan-gen-final",
        )
        self.receta_prep = Receta.objects.create(
            nombre="Preparacion Test",
            tipo=Receta.TIPO_PREPARACION,
            hash_contenido="hash-plan-gen-prep",
        )
        PronosticoVenta.objects.create(receta=self.receta_final, periodo="2026-03", cantidad=Decimal("12"))
        PronosticoVenta.objects.create(receta=self.receta_prep, periodo="2026-03", cantidad=Decimal("8"))

    def test_genera_plan_desde_pronostico_solo_producto_final_por_default(self):
        response = self.client.post(
            reverse("recetas:plan_produccion_generar_desde_pronostico"),
            {
                "periodo": "2026-03",
                "fecha_produccion": "2026-03-10",
                "nombre": "Plan desde pronostico test",
            },
        )
        self.assertEqual(response.status_code, 302)
        plan = PlanProduccion.objects.get(nombre="Plan desde pronostico test")
        items = list(plan.items.select_related("receta").all())
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0].receta_id, self.receta_final.id)
        self.assertEqual(items[0].cantidad, Decimal("12"))

    def test_genera_plan_desde_pronostico_incluyendo_preparaciones(self):
        response = self.client.post(
            reverse("recetas:plan_produccion_generar_desde_pronostico"),
            {
                "periodo": "2026-03",
                "fecha_produccion": "2026-03-12",
                "nombre": "Plan pronostico con preparaciones",
                "incluir_preparaciones": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        plan = PlanProduccion.objects.get(nombre="Plan pronostico con preparaciones")
        self.assertEqual(plan.items.count(), 2)


class PronosticoEstadisticoDesdeHistorialTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_hist_forecast",
            email="admin_hist_forecast@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)
        self.sucursal = Sucursal.objects.create(codigo="MATRIZ", nombre="Matriz", activa=True)
        self.receta = Receta.objects.create(
            nombre="Pastel Forecast",
            codigo_point="P-FC-01",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-forecast-hist-001",
        )
        self.point_branch = PointBranch.objects.create(
            external_id="PB-MATRIZ-TEST",
            name="Matriz",
            erp_branch=self.sucursal,
        )

    def _create_canonical_sale(self, *, fecha: date, cantidad: str | Decimal, sucursal: Sucursal | None = None, receta: Receta | None = None):
        suc = sucursal or self.sucursal
        rec = receta or self.receta
        branch = self.point_branch
        if suc.id != self.sucursal.id:
            branch, _ = PointBranch.objects.get_or_create(
                external_id=f"PB-{suc.codigo}",
                defaults={
                    "name": suc.nombre,
                    "erp_branch": suc,
                },
            )
        PointSalesDailyProductFact.objects.create(
            branch=branch,
            sale_date=fecha,
            sucursal_nombre=suc.nombre,
            categoria="Pasteles",
            producto_nombre_historico=rec.nombre,
            receta=rec,
            match_catalogo_status="EXACT_CODE",
            total_cantidad=Decimal(str(cantidad)),
            total_venta=Decimal("100.00"),
            total_venta_neta=Decimal("100.00"),
        )

    def test_import_ventas_historicas_crea_registros(self):
        csv_data = (
            "receta,codigo_point,sucursal_codigo,fecha,cantidad,tickets,monto_total\n"
            "Pastel Forecast,P-FC-01,MATRIZ,2026-01-10,12,8,1200\n"
            "Pastel Forecast,P-FC-01,MATRIZ,2026-01-11,9,6,900\n"
        ).encode("utf-8")
        upload = SimpleUploadedFile("ventas.csv", csv_data, content_type="text/csv")
        response = self.client.post(
            reverse("recetas:ventas_historicas_importar"),
            {"archivo": upload, "modo": "replace", "fuente": "TEST_VENTAS"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(VentaHistorica.objects.count(), 2)
        row = VentaHistorica.objects.order_by("fecha").first()
        self.assertEqual(row.sucursal_id, self.sucursal.id)
        self.assertEqual(row.receta_id, self.receta.id)
        self.assertEqual(row.fuente, "TEST_VENTAS")

    def test_pronostico_estadistico_crea_plan(self):
        base = date(2026, 3, 12)  # jueves
        # 8 semanas de historial con patrón estable por día.
        for w in range(1, 9):
            week_start = base - timedelta(days=base.weekday()) - timedelta(days=(7 * w))
            for d in range(7):
                self._create_canonical_sale(fecha=week_start + timedelta(days=d), cantidad="5")

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "semana",
                "fecha_base": "2026-03-12",
                "periodo": "2026-03",
                "sucursal_id": str(self.sucursal.id),
                "run_mode": "crear_plan",
                "safety_pct": "0",
            },
        )
        self.assertEqual(response.status_code, 302)
        plan = PlanProduccion.objects.order_by("-id").first()
        self.assertIsNotNone(plan)
        self.assertIn("pronóstico estadístico", plan.notas.lower())
        self.assertEqual(plan.items.count(), 1)
        self.assertGreater(plan.items.first().cantidad, Decimal("0"))

        preview = self.client.session.get("pronostico_estadistico_preview")
        self.assertIsNotNone(preview)
        self.assertEqual(preview["sucursal_nombre"], "MATRIZ - Matriz")
        self.assertGreaterEqual(preview["totals"]["recetas_count"], 1)
        self.assertIn("history_meta", preview)
        self.assertGreaterEqual(preview["history_meta"]["years_observed"], 1)

    def test_pronostico_estadistico_aplica_pronostico_mensual(self):
        for month_idx, qty in [(10, "40"), (11, "50"), (12, "60"), (1, "65"), (2, "70")]:
            year = 2025 if month_idx >= 10 else 2026
            self._create_canonical_sale(fecha=date(year, month_idx, 15), cantidad=qty)

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "mes",
                "periodo": "2026-03",
                "sucursal_id": str(self.sucursal.id),
                "run_mode": "apply_pronostico",
                "safety_pct": "0",
            },
        )
        self.assertEqual(response.status_code, 302)
        pron = PronosticoVenta.objects.filter(receta=self.receta, periodo="2026-03").first()
        self.assertIsNotNone(pron)
        self.assertGreater(pron.cantidad, Decimal("0"))

    def test_pronostico_estadistico_backtest_genera_contexto(self):
        for month_idx, qty in [(9, "30"), (10, "35"), (11, "40"), (12, "46"), (1, "50"), (2, "55"), (3, "60")]:
            year = 2025 if month_idx >= 9 else 2026
            self._create_canonical_sale(fecha=date(year, month_idx, 15), cantidad=qty)

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "mes",
                "periodo": "2026-04",
                "fecha_base": "2026-04-15",
                "sucursal_id": str(self.sucursal.id),
                "run_mode": "backtest",
                "escenario": "alto",
                "backtest_periods": "4",
                "backtest_top": "5",
                "safety_pct": "0",
            },
        )
        self.assertEqual(response.status_code, 302)
        preview = self.client.session.get("pronostico_backtest_preview")
        self.assertIsNotNone(preview)
        self.assertGreaterEqual(preview["totals"]["windows_evaluated"], 1)
        self.assertEqual(preview["scope"]["escenario"], "alto")
        self.assertIn("mape_promedio", preview["totals"])
        self.assertGreaterEqual(len(preview["windows"]), 1)
        self.assertIn("history_meta", preview)
        self.assertGreaterEqual(preview["history_meta"]["years_observed"], 1)

        response = self.client.get(reverse("recetas:plan_produccion"), {"periodo": "2026-04"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("forecast_backtest", response.context)
        self.assertIsNotNone(response.context["forecast_backtest"])
        self.assertContains(response, "Años observados")
        self.assertContains(response, "Temporadas comparables")

    def test_pronostico_estadistico_backtest_export_csv_y_xlsx(self):
        for month_idx, qty in [(9, "30"), (10, "35"), (11, "40"), (12, "46"), (1, "50"), (2, "55"), (3, "60")]:
            year = 2025 if month_idx >= 9 else 2026
            self._create_canonical_sale(fecha=date(year, month_idx, 15), cantidad=qty)

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "mes",
                "periodo": "2026-04",
                "fecha_base": "2026-04-15",
                "sucursal_id": str(self.sucursal.id),
                "run_mode": "backtest",
                "escenario": "alto",
                "backtest_periods": "4",
                "backtest_top": "5",
                "safety_pct": "0",
            },
        )
        self.assertEqual(response.status_code, 302)

        response_csv = self.client.get(reverse("recetas:forecast_backtest_export"), {"format": "csv", "periodo": "2026-04"})
        self.assertEqual(response_csv.status_code, 200)
        self.assertIn("text/csv", response_csv["Content-Type"])
        body_csv = response_csv.content.decode("utf-8")
        self.assertIn("VENTANAS", body_csv)
        self.assertIn("TOP_ERRORES", body_csv)
        self.assertIn("Pastel Forecast", body_csv)

        response_xlsx = self.client.get(reverse("recetas:forecast_backtest_export"), {"format": "xlsx", "periodo": "2026-04"})
        self.assertEqual(response_xlsx.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response_xlsx["Content-Type"],
        )
        wb = load_workbook(BytesIO(response_xlsx.content), data_only=True)
        self.assertIn("Resumen", wb.sheetnames)
        self.assertIn("Ventanas", wb.sheetnames)
        self.assertIn("TopErrores", wb.sheetnames)

    def test_pronostico_estadistico_min_confianza_filtra_resultados(self):
        for month_idx, qty in [(10, "40"), (11, "50"), (12, "60"), (1, "65"), (2, "70")]:
            year = 2025 if month_idx >= 10 else 2026
            self._create_canonical_sale(fecha=date(year, month_idx, 15), cantidad=qty)

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "mes",
                "periodo": "2026-03",
                "sucursal_id": str(self.sucursal.id),
                "run_mode": "preview",
                "safety_pct": "0",
                "min_confianza_pct": "90",
            },
        )
        self.assertEqual(response.status_code, 302)
        preview = self.client.session.get("pronostico_estadistico_preview")
        self.assertIsNotNone(preview)
        self.assertEqual(preview["totals"]["recetas_count"], 0)
        self.assertEqual(len(preview["rows"]), 0)

    def test_pronostico_estadistico_preview_export_csv_y_xlsx(self):
        for month_idx, qty in [(10, "40"), (11, "50"), (12, "60"), (1, "65"), (2, "70")]:
            year = 2025 if month_idx >= 10 else 2026
            self._create_canonical_sale(fecha=date(year, month_idx, 15), cantidad=qty)

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "mes",
                "periodo": "2026-03",
                "sucursal_id": str(self.sucursal.id),
                "run_mode": "preview",
                "safety_pct": "0",
            },
        )
        self.assertEqual(response.status_code, 302)

        response_csv = self.client.get(reverse("recetas:forecast_preview_export"), {"format": "csv", "periodo": "2026-03"})
        self.assertEqual(response_csv.status_code, 200)
        self.assertIn("text/csv", response_csv["Content-Type"])
        body_csv = response_csv.content.decode("utf-8")
        self.assertIn("receta_id,receta,forecast,banda_baja,banda_alta,pronostico_actual", body_csv)
        self.assertIn("Pastel Forecast", body_csv)

        response_xlsx = self.client.get(reverse("recetas:forecast_preview_export"), {"format": "xlsx", "periodo": "2026-03"})
        self.assertEqual(response_xlsx.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response_xlsx["Content-Type"],
        )
        wb = load_workbook(BytesIO(response_xlsx.content), data_only=True)
        self.assertIn("Resumen", wb.sheetnames)
        self.assertIn("Detalle", wb.sheetnames)

    def test_pronostico_estadistico_dia_generates_branch_detail_rows(self):
        sucursal_b = Sucursal.objects.create(codigo="NORTE", nombre="Norte", activa=True)
        target_day = date(2026, 4, 6)  # lunes
        for lag in range(1, 9):
            sample_day = target_day - timedelta(days=7 * lag)
            self._create_canonical_sale(fecha=sample_day, cantidad="12")
            self._create_canonical_sale(fecha=sample_day, cantidad="5", sucursal=sucursal_b)

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "dia",
                "fecha_base": "2026-04-06",
                "run_mode": "preview",
                "safety_pct": "0",
            },
        )
        self.assertEqual(response.status_code, 302)
        preview = self.client.session.get("pronostico_estadistico_preview")
        self.assertIsNotNone(preview)
        self.assertEqual(preview["alcance"], "dia")
        self.assertGreaterEqual(preview["detail_rows_total"], 2)
        detail_rows = preview["detail_rows"]
        self.assertTrue(any(row["sucursal_codigo"] == "MATRIZ" for row in detail_rows))
        self.assertTrue(any(row["sucursal_codigo"] == "NORTE" for row in detail_rows))
        self.assertTrue(all(row["fecha"] == "2026-04-06" for row in detail_rows))
        self.assertIn("model_meta", preview)
        self.assertEqual(preview["model_meta"]["history_selection"]["history_start"], "2022-01-01")
        self.assertFalse(preview["model_meta"]["mix_adjuster"]["enabled"])

    def test_pronostico_estadistico_mix_adjuster_caps_changes(self):
        receta_b = Receta.objects.create(
            nombre="Pastel Forecast Mix B",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-forecast-mix-b",
        )
        target_day = date(2026, 4, 6)
        for lag in range(1, 10):
            sample_day = target_day - timedelta(days=7 * lag)
            self._create_canonical_sale(fecha=sample_day, cantidad="10", receta=self.receta)
            self._create_canonical_sale(fecha=sample_day, cantidad="10", receta=receta_b)
        for offset in (5, 12, 19):
            recent_day = target_day - timedelta(days=offset)
            self._create_canonical_sale(fecha=recent_day, cantidad="18", receta=self.receta)
            self._create_canonical_sale(fecha=recent_day, cantidad="2", receta=receta_b)

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "dia",
                "fecha_base": target_day.isoformat(),
                "sucursal_id": str(self.sucursal.id),
                "run_mode": "preview",
                "mix_adjustment_enabled": "1",
                "safety_pct": "0",
            },
        )
        self.assertEqual(response.status_code, 302)
        preview = self.client.session.get("pronostico_estadistico_preview")
        self.assertIsNotNone(preview)
        self.assertTrue(preview["model_meta"]["mix_adjuster"]["enabled"])
        detail_rows = {row["receta_id"]: row for row in preview["detail_rows"]}
        qty_a = Decimal(str(detail_rows[self.receta.id]["forecast_qty"]))
        qty_b = Decimal(str(detail_rows[receta_b.id]["forecast_qty"]))
        self.assertGreater(qty_a, Decimal("0"))
        self.assertGreater(qty_b, Decimal("0"))
        self.assertTrue(detail_rows[self.receta.id]["mix_adjusted"])
        self.assertLessEqual(
            Decimal(str(detail_rows[self.receta.id]["mix_adjustment_pct"])),
            Decimal("20.0"),
        )
        self.assertGreaterEqual(
            Decimal(str(detail_rows[receta_b.id]["mix_adjustment_pct"])),
            Decimal("-20.0"),
        )

    def test_pronostico_estadistico_downweights_march_2026_legacy_snapshot(self):
        target_day = date(2027, 3, 10)
        comparable_days = [
            (date(2024, 3, 13), "20"),
            (date(2025, 3, 12), "22"),
            (date(2026, 3, 11), "250"),
        ]
        for sample_day, qty in comparable_days:
            self._create_canonical_sale(fecha=sample_day, cantidad=qty)

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "dia",
                "fecha_base": target_day.isoformat(),
                "sucursal_id": str(self.sucursal.id),
                "run_mode": "preview",
                "safety_pct": "0",
            },
        )
        self.assertEqual(response.status_code, 302)
        preview = self.client.session.get("pronostico_estadistico_preview")
        self.assertIsNotNone(preview)
        self.assertGreaterEqual(len(preview["rows"]), 1)
        row = preview["rows"][0]
        self.assertLess(row["forecast_qty"], 80.0)
        self.assertIn("model_meta", preview)
        adjustment = preview["model_meta"]["data_quality_adjustments"][0]
        self.assertEqual(adjustment["period"], "2026-03")
        self.assertEqual(adjustment["action"], "excluded_from_baseline")

    def test_pronostico_estadistico_does_not_fallback_to_venta_historica_raw(self):
        target_day = date(2026, 4, 6)
        for lag in range(1, 9):
            sample_day = target_day - timedelta(days=7 * lag)
            VentaHistorica.objects.create(
                receta=self.receta,
                sucursal=self.sucursal,
                fecha=sample_day,
                cantidad=Decimal("25"),
                fuente="RAW_ONLY",
            )

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "dia",
                "fecha_base": target_day.isoformat(),
                "sucursal_id": str(self.sucursal.id),
                "run_mode": "preview",
                "safety_pct": "0",
            },
        )
        self.assertEqual(response.status_code, 302)
        preview = self.client.session.get("pronostico_estadistico_preview")
        self.assertIsNotNone(preview)
        self.assertEqual(preview["totals"]["recetas_count"], 0)

    def test_evaluar_mix_adjuster_command_generates_comparison_artifacts(self):
        receta_b = Receta.objects.create(
            nombre="Pastel Forecast Eval Mix B",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-forecast-eval-mix-b",
        )
        target_day = date(2026, 4, 6)
        for lag in range(1, 10):
            sample_day = target_day - timedelta(days=7 * lag)
            self._create_canonical_sale(fecha=sample_day, cantidad="10", receta=self.receta)
            self._create_canonical_sale(fecha=sample_day, cantidad="10", receta=receta_b)
        for offset in (5, 12, 19):
            recent_day = target_day - timedelta(days=offset)
            self._create_canonical_sale(fecha=recent_day, cantidad="18", receta=self.receta)
            self._create_canonical_sale(fecha=recent_day, cantidad="2", receta=receta_b)

        with tempfile.TemporaryDirectory() as tmpdir:
            call_command(
                "evaluar_mix_adjuster",
                "--fecha-base",
                target_day.isoformat(),
                "--alcance",
                "dia",
                "--periods",
                "4",
                "--sucursal-id",
                str(self.sucursal.id),
                "--output-dir",
                tmpdir,
                "--label",
                "test",
            )
            run_dirs = [name for name in os.listdir(tmpdir) if name.startswith("mix_adjuster_eval_")]
            self.assertEqual(len(run_dirs), 1)
            run_dir = os.path.join(tmpdir, run_dirs[0])
            summary_path = os.path.join(run_dir, "summary.csv")
            windows_path = os.path.join(run_dir, "windows.csv")
            cases_path = os.path.join(run_dir, "cases.csv")
            json_path = os.path.join(run_dir, "summary.json")
            self.assertTrue(os.path.exists(summary_path))
            self.assertTrue(os.path.exists(windows_path))
            self.assertTrue(os.path.exists(cases_path))
            self.assertTrue(os.path.exists(json_path))
            with open(summary_path, "r", encoding="utf-8") as fh:
                summary_csv = fh.read()
            self.assertIn("base_mape", summary_csv)
            self.assertIn("adjusted_mape", summary_csv)
            self.assertIn("recommendation", summary_csv)

    def test_evaluar_mix_adjuster_all_sucursales_excludes_col_duplicate(self):
        Sucursal.objects.create(codigo="COL", nombre="Colosio", activa=True)
        with tempfile.TemporaryDirectory() as tmpdir:
            call_command(
                "evaluar_mix_adjuster",
                "--fecha-base",
                date(2026, 4, 6).isoformat(),
                "--alcance",
                "dia",
                "--periods",
                "2",
                "--all-sucursales",
                "--output-dir",
                tmpdir,
                "--label",
                "exclude_col",
            )
            run_dirs = [name for name in os.listdir(tmpdir) if name.startswith("mix_adjuster_eval_")]
            self.assertEqual(len(run_dirs), 1)
            summary_path = os.path.join(tmpdir, run_dirs[0], "summary.csv")
            with open(summary_path, "r", encoding="utf-8") as fh:
                summary_csv = fh.read()
            self.assertNotIn("COL - Colosio", summary_csv)


class ImportarVentasPointBranchResolutionTests(TestCase):
    def setUp(self):
        self.command = ImportarVentasPointArchivosCommand()

    def test_resolve_default_sucursal_rejects_excluded_duplicate_code(self):
        Sucursal.objects.create(codigo="COLOSIO", nombre="Colosio", activa=True)
        Sucursal.objects.create(codigo="COL", nombre="Colosio", activa=True)

        with self.assertRaises(CommandError):
            self.command._resolve_default_sucursal("COL")

    def test_resolve_sucursal_uses_canonical_branch_for_duplicate_name(self):
        canonical = Sucursal.objects.create(codigo="COLOSIO", nombre="Colosio", activa=True)
        Sucursal.objects.create(codigo="COL", nombre="Colosio", activa=True)

        resolved = self.command._resolve_sucursal(
            sucursal_name="Colosio",
            sucursal_code="",
            default_sucursal=None,
            cache={},
        )

        self.assertIsNotNone(resolved)
        self.assertEqual(resolved.id, canonical.id)


class SolicitudVentasForecastTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_sol_ventas",
            email="admin_sol_ventas@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)
        self.sucursal = Sucursal.objects.create(codigo="MATRIZ", nombre="Matriz", activa=True)
        self.receta = Receta.objects.create(
            nombre="Pastel Solicitud",
            codigo_point="P-SOL-01",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-sol-ventas-001",
        )

    def _crear_producto_plantilla(
        self,
        *,
        nombre,
        codigo,
        familia,
        modo_costeo=Receta.MODO_COSTEO_FABRICADO,
        alias_activo=True,
        point_activo=True,
        crear_point=True,
        codigo_receta=None,
        vendido=True,
        categoria_point=None,
    ):
        receta = Receta.objects.create(
            nombre=nombre,
            codigo_point=codigo if codigo_receta is None else codigo_receta,
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            modo_costeo=modo_costeo,
            familia=familia,
            hash_contenido=f"hash-plantilla-{codigo.lower()}",
        )
        RecetaCodigoPointAlias.objects.create(
            receta=receta,
            codigo_point=codigo,
            nombre_point=nombre,
            activo=alias_activo,
        )
        if crear_point:
            point_product = PointProduct.objects.create(
                external_id=f"point-{codigo.lower()}",
                sku=codigo,
                name=nombre,
                category=categoria_point or familia,
                active=point_activo,
            )
            if vendido:
                point_branch, _ = PointBranch.objects.get_or_create(
                    external_id="POINT-PLANTILLA",
                    defaults={"name": "Point Plantilla"},
                )
                PointDailySale.objects.create(
                    branch=point_branch,
                    product=point_product,
                    receta=receta,
                    sale_date=timezone.localdate(),
                    quantity=Decimal("1"),
                )
        return receta

    def test_calculo_insumos_plantilla_selecciona_productos_vendibles(self):
        from recetas.views.plan import _calculo_insumos_template_products

        self._crear_producto_plantilla(
            nombre="Pastel Vendible",
            codigo="PASTEL-01",
            familia="Pastel",
            categoria_point="Pastel Mediano",
        )
        self._crear_producto_plantilla(
            nombre="Vela Número",
            codigo="VELA-01",
            familia="Velas Sparklers",
        )
        self._crear_producto_plantilla(
            nombre="Letrero Feliz",
            codigo="LET-01",
            familia="Letrero B",
        )
        self._crear_producto_plantilla(
            nombre="Kit Repostería",
            codigo="ACC-01",
            familia="Accesorios de repostería",
        )
        self._crear_producto_plantilla(
            nombre="Tarjeta de regalo",
            codigo="REGALO-01",
            familia="Otros postres",
        )
        self._crear_producto_plantilla(
            nombre="Capuchino",
            codigo="CAFE-01",
            familia="Bebidas",
            categoria_point="Café",
        )
        self._crear_producto_plantilla(
            nombre="Servicio Accesorio",
            codigo="SERV-01",
            familia="Pasteles",
            modo_costeo=Receta.MODO_COSTEO_SERVICIO,
        )
        self._crear_producto_plantilla(
            nombre="Pastel Inactivo",
            codigo="PASTEL-OFF",
            familia="Pasteles",
            alias_activo=False,
            codigo_receta="",
        )
        self._crear_producto_plantilla(
            nombre="Pastel Fuera de Point",
            codigo="PASTEL-NO-POINT",
            familia="Pasteles",
            crear_point=False,
        )
        self._crear_producto_plantilla(
            nombre="Pastel Sin Venta",
            codigo="PASTEL-SIN-VENTA",
            familia="Pastel",
            categoria_point="Pastel Grande",
            vendido=False,
        )

        self.assertEqual(
            _calculo_insumos_template_products(),
            [
                {
                    "familia": "Pastel",
                    "categoria": "Mediano",
                    "codigo_point": "PASTEL-01",
                    "producto": "Pastel Vendible",
                }
            ],
        )

    def test_calculo_insumos_plantilla_xlsx_precarga_catalogo_vendible(self):
        self._crear_producto_plantilla(
            nombre="Pastel Vendible",
            codigo="PASTEL-01",
            familia="Pastel",
            categoria_point="Pastel Mediano",
        )
        self._crear_producto_plantilla(
            nombre="Vela Número",
            codigo="VELA-01",
            familia="Velas Sparklers",
        )
        self._crear_producto_plantilla(
            nombre="Letrero Feliz",
            codigo="LET-01",
            familia="Letrero B",
        )

        response = self.client.get(reverse("recetas:calculo_insumos_plantilla"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Cache-Control"], "no-store, no-cache, must-revalidate, max-age=0")
        self.assertEqual(
            response["Content-Disposition"],
            'attachment; filename="plantilla_calculo_insumos.xlsx"',
        )
        wb = load_workbook(BytesIO(response.content), data_only=True)
        ws = wb["Plantilla carga"]
        self.assertEqual(
            list(ws.values),
            [
                ("familia", "categoria", "codigo_point", "producto", "cantidad", "notas"),
                ("Pastel", "Mediano", "PASTEL-01", "Pastel Vendible", None, None),
            ],
        )

    def test_guardar_solicitud_venta_reemplaza_por_rango(self):
        payload = {
            "receta_id": str(self.receta.id),
            "sucursal_id": str(self.sucursal.id),
            "alcance": "mes",
            "periodo": "2026-04",
            "cantidad": "120",
            "fuente": "TEST_SOL",
        }
        response = self.client.post(reverse("recetas:solicitud_ventas_guardar"), payload)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(SolicitudVenta.objects.count(), 1)
        row = SolicitudVenta.objects.first()
        self.assertEqual(row.periodo, "2026-04")
        self.assertEqual(row.cantidad, Decimal("120"))

        payload["cantidad"] = "145"
        response = self.client.post(reverse("recetas:solicitud_ventas_guardar"), payload)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(SolicitudVenta.objects.count(), 1)
        row.refresh_from_db()
        self.assertEqual(row.cantidad, Decimal("145"))

    def test_import_solicitud_ventas_csv_crea_registros(self):
        csv_data = (
            "codigo_point,sucursal_codigo,alcance,periodo,cantidad\n"
            "P-SOL-01,MATRIZ,MES,2026-04,88\n"
            "P-SOL-01,MATRIZ,SEMANA,2026-04,24\n"
        ).encode("utf-8")
        upload = SimpleUploadedFile("solicitudes_ventas.csv", csv_data, content_type="text/csv")
        response = self.client.post(
            reverse("recetas:solicitud_ventas_importar"),
            {
                "archivo": upload,
                "modo": "replace",
                "periodo_default": "2026-04",
                "fecha_base_default": "2026-04-10",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(SolicitudVenta.objects.count(), 2)

    def test_contexto_muestra_comparativo_pronostico_vs_solicitud(self):
        for month_idx, qty in [(11, "60"), (12, "72"), (1, "81"), (2, "78"), (3, "90")]:
            year = 2025 if month_idx >= 11 else 2026
            VentaHistorica.objects.create(
                receta=self.receta,
                sucursal=self.sucursal,
                fecha=date(year, month_idx, 15),
                cantidad=Decimal(qty),
                fuente="TEST_HIST_SOL",
            )

        SolicitudVenta.objects.create(
            receta=self.receta,
            sucursal=self.sucursal,
            alcance=SolicitudVenta.ALCANCE_MES,
            periodo="2026-04",
            fecha_inicio=date(2026, 4, 1),
            fecha_fin=date(2026, 4, 30),
            cantidad=Decimal("110"),
            fuente="TEST_SOL",
        )

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "mes",
                "periodo": "2026-04",
                "sucursal_id": str(self.sucursal.id),
                "run_mode": "preview",
                "safety_pct": "0",
            },
        )
        self.assertEqual(response.status_code, 302)

        response = self.client.get(reverse("recetas:plan_produccion"), {"periodo": "2026-04"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("forecast_vs_solicitud", response.context)
        compare = response.context["forecast_vs_solicitud"]
        self.assertIsNotNone(compare)
        self.assertGreater(len(compare["rows"]), 0)
        row = next((r for r in compare["rows"] if r["receta_id"] == self.receta.id), None)
        self.assertIsNotNone(row)
        self.assertEqual(row["solicitud_qty"], Decimal("110"))
        self.assertEqual(row["forecast_qty"], row["forecast_base"])

    def test_plan_produccion_muestra_explosion_de_insumos_desde_preview_forecast(self):
        unidad = UnidadMedida.objects.create(codigo="kg-fsp", nombre="Kg forecast supply", tipo=UnidadMedida.TIPO_MASA)
        proveedor = Proveedor.objects.create(nombre="Proveedor Forecast Supply")
        harina = Insumo.objects.create(
            nombre="Harina Forecast Supply",
            nombre_normalizado="harina forecast supply",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
        )
        CostoInsumo.objects.create(
            insumo=harina,
            proveedor=proveedor,
            fecha=date(2026, 4, 1),
            costo_unitario=Decimal("8.00"),
            source_hash="forecast-supply-view-cost",
        )
        LineaReceta.objects.create(
            receta=self.receta,
            posicion=1,
            insumo=harina,
            insumo_texto=harina.nombre,
            cantidad=Decimal("2.5"),
            unidad=unidad,
            unidad_texto="kg",
            match_status=LineaReceta.STATUS_AUTO,
        )
        session = self.client.session
        session["pronostico_estadistico_preview"] = {
            "target_start": "2026-04-01",
            "target_end": "2026-04-30",
            "target_label": "Abril 2026",
            "alcance": "mes",
            "periodo": "2026-04",
            "sucursal_nombre": "MATRIZ - Matriz",
            "totals": {"forecast_total": 10.0, "recetas_count": 1},
            "rows": [
                {
                    "receta_id": self.receta.id,
                    "receta": self.receta.nombre,
                    "forecast_qty": 10.0,
                    "forecast_low": 8.0,
                    "forecast_high": 12.0,
                }
            ],
        }
        session.save()

        response = self.client.get(reverse("recetas:plan_produccion"), {"periodo": "2026-04"})
        self.assertEqual(response.status_code, 200)
        supply = response.context["forecast_supply_context"]
        self.assertIsNotNone(supply)
        self.assertEqual(supply["summary"]["estimated_spend"], Decimal("200.00"))
        self.assertContains(response, "Insumos requeridos desde pronóstico")
        self.assertContains(response, "Materia prima y empaques")
        self.assertContains(response, "Harina Forecast Supply")
        self.assertContains(response, "$200.00")

    def test_forecast_supply_exporta_xlsx_profesional(self):
        unidad = UnidadMedida.objects.create(codigo="kg-fsx", nombre="Kg forecast supply xlsx", tipo=UnidadMedida.TIPO_MASA)
        proveedor = Proveedor.objects.create(nombre="Proveedor Forecast Supply XLSX")
        harina = Insumo.objects.create(
            nombre="Harina Forecast Supply XLSX",
            nombre_normalizado="harina forecast supply xlsx",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            categoria="Harinas",
        )
        CostoInsumo.objects.create(
            insumo=harina,
            proveedor=proveedor,
            fecha=date(2026, 4, 1),
            costo_unitario=Decimal("10.00"),
            source_hash="forecast-supply-xlsx-cost",
        )
        LineaReceta.objects.create(
            receta=self.receta,
            posicion=1,
            insumo=harina,
            insumo_texto=harina.nombre,
            cantidad=Decimal("3"),
            unidad=unidad,
            unidad_texto="kg",
            match_status=LineaReceta.STATUS_AUTO,
        )
        session = self.client.session
        session["pronostico_estadistico_preview"] = {
            "target_start": "2026-04-01",
            "target_end": "2026-04-30",
            "target_label": "Abril 2026",
            "alcance": "mes",
            "periodo": "2026-04",
            "sucursal_nombre": "MATRIZ - Matriz",
            "totals": {"forecast_total": 4.0, "recetas_count": 1},
            "rows": [
                {
                    "receta_id": self.receta.id,
                    "receta": self.receta.nombre,
                    "forecast_qty": 4.0,
                    "forecast_low": 3.0,
                    "forecast_high": 5.0,
                }
            ],
        }
        session.save()

        response = self.client.get(reverse("recetas:forecast_supply_export"), {"escenario": "base"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        wb = load_workbook(BytesIO(response.content), data_only=True)
        self.assertEqual(
            wb.sheetnames,
            [
                "Resumen",
                "Productos forecast",
                "Preparados a producir",
                "Materia prima y empaques",
                "Explosion multinivel",
                "Faltantes por corregir",
            ],
        )
        ws = wb["Materia prima y empaques"]
        self.assertEqual(ws["D2"].value, "Harina Forecast Supply XLSX")
        self.assertEqual(ws["E2"].value, 12)
        self.assertEqual(ws["H2"].value, 120)

    def test_calculo_insumos_importa_csv_y_muestra_bom(self):
        unidad = UnidadMedida.objects.create(codigo="g", nombre="Gramo cálculo insumos", tipo=UnidadMedida.TIPO_MASA)
        proveedor = Proveedor.objects.create(nombre="Proveedor Cálculo Insumos")
        azucar = Insumo.objects.create(
            nombre="Azúcar Cálculo Insumos",
            codigo_point="AZ-CALC-01",
            nombre_normalizado="azucar calculo insumos",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
        )
        CostoInsumo.objects.create(
            insumo=azucar,
            proveedor=proveedor,
            fecha=date(2026, 4, 1),
            costo_unitario=Decimal("0.02"),
            source_hash="calculo-insumos-cost",
        )
        LineaReceta.objects.create(
            receta=self.receta,
            posicion=1,
            insumo=azucar,
            insumo_texto=azucar.nombre,
            cantidad=Decimal("500"),
            unidad=unidad,
            unidad_texto="g",
            match_status=LineaReceta.STATUS_AUTO,
        )
        csv_data = (
            "pastel,presentacion,cantidad\n"
            "Pastel Solicitud,,4\n"
        ).encode("utf-8")
        upload = SimpleUploadedFile("calculo_insumos.csv", csv_data, content_type="text/csv")

        response = self.client.post(reverse("recetas:calculo_insumos_importar"), {"archivo": upload})
        self.assertEqual(response.status_code, 302)
        response = self.client.get(reverse("recetas:plan_produccion"), {"seccion": "calculo_insumos"})
        self.assertEqual(response.status_code, 200)
        context = response.context["calculo_insumos_context"]
        self.assertIsNotNone(context)
        supply = context["supply"]
        self.assertEqual(supply["summary"]["estimated_spend"], Decimal("40.00"))
        row = supply["insumos"][0]
        self.assertEqual(row["codigo_point"], "AZ-CALC-01")
        self.assertEqual(row["display_required_qty"], Decimal("2.000"))
        self.assertEqual(row["display_unidad_codigo"], "kg")
        self.assertContains(response, "Cálculo de insumos")
        self.assertContains(response, "Azúcar Cálculo Insumos")
        self.assertContains(response, "$40.00")

    def test_calculo_insumos_importa_csv_excel_espanol_con_miles(self):
        unidad = UnidadMedida.objects.create(codigo="g-miles", nombre="Gramo cálculo miles", tipo=UnidadMedida.TIPO_MASA)
        proveedor = Proveedor.objects.create(nombre="Proveedor Cálculo Miles")
        azucar = Insumo.objects.create(
            nombre="Azúcar Cálculo Miles",
            nombre_normalizado="azucar calculo miles",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
        )
        CostoInsumo.objects.create(
            insumo=azucar,
            proveedor=proveedor,
            fecha=date(2026, 4, 1),
            costo_unitario=Decimal("0.01"),
            source_hash="calculo-insumos-miles-cost",
        )
        LineaReceta.objects.create(
            receta=self.receta,
            posicion=1,
            insumo=azucar,
            insumo_texto=azucar.nombre,
            cantidad=Decimal("1"),
            unidad=unidad,
            unidad_texto="g",
            match_status=LineaReceta.STATUS_AUTO,
        )
        csv_data = (
            "codigo_point;producto;presentacion;cantidad\n"
            "P-SOL-01;Pastel Solicitud;;1,000\n"
        ).encode("utf-8")
        upload = SimpleUploadedFile("calculo_insumos.csv", csv_data, content_type="text/csv")

        response = self.client.post(reverse("recetas:calculo_insumos_importar"), {"archivo": upload})
        self.assertEqual(response.status_code, 302)
        response = self.client.get(reverse("recetas:plan_produccion"), {"seccion": "calculo_insumos"})
        context = response.context["calculo_insumos_context"]
        self.assertEqual(context["preview"]["rows"][0]["forecast_qty"], 1000.0)
        self.assertEqual(context["supply"]["insumos"][0]["required_gross_qty"], Decimal("1000.000"))

    def test_calculo_insumos_captura_manual_respeta_decimal_con_punto(self):
        response = self.client.post(
            reverse("recetas:calculo_insumos_guardar"),
            {
                "receta_id": str(self.receta.id),
                "cantidad": "1.000",
                "notas": "decimal manual",
            },
        )

        self.assertEqual(response.status_code, 302)
        session = self.client.session
        self.assertNotIn("calculo_insumos_preview", session)
        draft = session["calculo_insumos_draft"]
        self.assertEqual(draft["source_rows"][0]["cantidad"], "1.000")
        self.assertEqual(draft["source_rows"][0]["producto"], "Pastel Solicitud")

        response = self.client.post(reverse("recetas:calculo_insumos_calcular"))
        self.assertEqual(response.status_code, 302)
        session = self.client.session
        preview = session["calculo_insumos_preview"]
        self.assertEqual(preview["rows"][0]["forecast_qty"], 1.0)
        self.assertEqual(preview["source_rows"][0]["cantidad"], "1.000")

    def test_calculo_insumos_muestra_lista_manual_antes_de_calcular(self):
        self.client.post(
            reverse("recetas:calculo_insumos_guardar"),
            {
                "receta_id": str(self.receta.id),
                "cantidad": "3.000",
                "notas": "lista grande",
            },
        )

        response = self.client.get(reverse("recetas:plan_produccion"), {"seccion": "calculo_insumos"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Borrador manual por calcular")
        self.assertContains(response, "Calcular lista")
        self.assertNotContains(response, "Azúcar Cálculo Insumos")
        self.assertIsNone(response.context["calculo_insumos_context"]["supply"])

    def test_calculo_insumos_no_mezcla_planes_recientes_ni_detalle_operacion(self):
        plan = PlanProduccion.objects.create(
            nombre="Plan visible en operación",
            fecha_produccion=date(2026, 4, 2),
            creado_por=self.user,
        )
        PlanProduccionItem.objects.create(plan=plan, receta=self.receta, cantidad=Decimal("5"))

        response = self.client.get(
            reverse("recetas:plan_produccion"),
            {"seccion": "calculo_insumos", "plan_id": str(plan.id)},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cálculo de insumos")
        self.assertNotContains(response, "Planes recientes")
        self.assertNotContains(response, "Productos del plan")
        self.assertNotContains(response, "Plan visible en operación")

    def test_calculo_insumos_usa_proyeccion_actual_por_escenario(self):
        session = self.client.session
        session["pronostico_estadistico_preview"] = {
            "target_start": "2026-04-01",
            "target_end": "2026-04-30",
            "target_label": "Abril 2026",
            "alcance": "mes",
            "periodo": "2026-04",
            "sucursal_nombre": "MATRIZ - Matriz",
            "escenario": "base",
            "totals": {
                "forecast_total": 10.0,
                "forecast_low_total": 8.0,
                "forecast_high_total": 12.0,
                "recetas_count": 1,
            },
            "rows": [
                {
                    "receta_id": self.receta.id,
                    "receta": self.receta.nombre,
                    "codigo_point": self.receta.codigo_point,
                    "forecast_qty": 10.0,
                    "forecast_low": 8.0,
                    "forecast_high": 12.0,
                }
            ],
        }
        session.save()

        response = self.client.post(reverse("recetas:calculo_insumos_desde_proyeccion"), {"escenario": "alto"})

        self.assertEqual(response.status_code, 302)
        preview = self.client.session["calculo_insumos_preview"]
        self.assertEqual(preview["source_label"], "Proyección ALTO · Abril 2026")
        self.assertEqual(preview["rows"][0]["forecast_qty"], 12.0)
        self.assertEqual(preview["source_rows"][0]["codigo_point"], "P-SOL-01")

    def test_calculo_insumos_muestra_puente_con_proyeccion_activa(self):
        session = self.client.session
        session["pronostico_estadistico_preview"] = {
            "target_start": "2026-04-01",
            "target_end": "2026-04-30",
            "target_label": "Abril 2026",
            "alcance": "mes",
            "periodo": "2026-04",
            "sucursal_nombre": "MATRIZ - Matriz",
            "escenario": "base",
            "totals": {"forecast_total": 10.0, "recetas_count": 1},
            "rows": [
                {
                    "receta_id": self.receta.id,
                    "receta": self.receta.nombre,
                    "codigo_point": self.receta.codigo_point,
                    "forecast_qty": 10.0,
                    "forecast_low": 8.0,
                    "forecast_high": 12.0,
                }
            ],
        }
        session.save()

        response = self.client.get(reverse("recetas:plan_produccion"), {"seccion": "calculo_insumos"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Proyección activa")
        self.assertContains(response, "Usar proyección actual")

    def test_calculo_insumos_rechaza_preparacion_interna_en_captura(self):
        preparacion = Receta.objects.create(
            nombre="Mermelada Captura Directa",
            codigo_point="PREP-CALC-01",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("9.5"),
            hash_contenido="hash-prep-calc-directa",
        )

        response = self.client.get(reverse("recetas:plan_produccion"), {"seccion": "calculo_insumos"})
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Mermelada Captura Directa")

        csv_data = (
            "codigo_point,producto,cantidad\n"
            "PREP-CALC-01,Mermelada Captura Directa,1\n"
        ).encode("utf-8")
        upload = SimpleUploadedFile("calculo_insumos.csv", csv_data, content_type="text/csv")
        response = self.client.post(reverse("recetas:calculo_insumos_importar"), {"archivo": upload})
        self.assertEqual(response.status_code, 302)
        response = self.client.get(reverse("recetas:plan_produccion"), {"seccion": "calculo_insumos"})
        context = response.context["calculo_insumos_context"]
        self.assertEqual(len(context["unresolved_rows"]), 1)
        self.assertIn("producto final", context["unresolved_rows"][0]["motivo"])
        self.assertNotEqual(preparacion.tipo, Receta.TIPO_PRODUCTO_FINAL)

    def test_calculo_insumos_oculta_acciones_mutables_sin_permiso_change(self):
        readonly = get_user_model().objects.create_user(
            username="readonly_calc_insumos",
            email="readonly_calc_insumos@example.com",
            password="test12345",
        )
        readonly.user_permissions.add(Permission.objects.get(content_type__app_label="recetas", codename="view_planproduccion"))
        self.client.force_login(readonly)

        response = self.client.get(reverse("recetas:plan_produccion"), {"seccion": "calculo_insumos"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Modo consulta")
        self.assertNotContains(response, "Agregar a lista")
        self.assertNotContains(response, "Calcular lista")
        self.assertNotContains(response, "Importar y calcular")

    def test_calculo_insumos_exporta_xlsx_con_reglas_y_unidades_visibles(self):
        unidad = UnidadMedida.objects.create(codigo="g", nombre="Gramo cálculo XLSX", tipo=UnidadMedida.TIPO_MASA)
        proveedor = Proveedor.objects.create(nombre="Proveedor Cálculo XLSX")
        azucar = Insumo.objects.create(
            nombre="Azúcar Cálculo XLSX",
            codigo_point="AZ-XLSX-01",
            nombre_normalizado="azucar calculo xlsx",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
        )
        CostoInsumo.objects.create(
            insumo=azucar,
            proveedor=proveedor,
            fecha=date(2026, 4, 1),
            costo_unitario=Decimal("0.03"),
            source_hash="calculo-insumos-xlsx-cost",
        )
        LineaReceta.objects.create(
            receta=self.receta,
            posicion=1,
            insumo=azucar,
            insumo_texto=azucar.nombre,
            cantidad=Decimal("250"),
            unidad=unidad,
            unidad_texto="g",
            match_status=LineaReceta.STATUS_AUTO,
        )
        session = self.client.session
        session["calculo_insumos_preview"] = {
            "target_start": "2026-04-01",
            "target_end": "2026-04-01",
            "target_label": "Cálculo prueba",
            "alcance": "calculo_insumos",
            "periodo": "2026-04",
            "sucursal_nombre": "Producción",
            "source_label": "Prueba",
            "source_rows": [],
            "unresolved_rows": [],
            "skipped_rows": [],
            "totals": {"forecast_total": 8.0, "recetas_count": 1},
            "rows": [
                {
                    "receta_id": self.receta.id,
                    "receta": self.receta.nombre,
                    "codigo_point": self.receta.codigo_point,
                    "forecast_qty": 8.0,
                    "forecast_low": 8.0,
                    "forecast_high": 8.0,
                }
            ],
        }
        session.save()

        response = self.client.get(reverse("recetas:calculo_insumos_export"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        wb = load_workbook(BytesIO(response.content), data_only=True)
        self.assertIn("Resumen", wb.sheetnames)
        self.assertIn("BOM compras", wb.sheetnames)
        self.assertIn("Detalle producción", wb.sheetnames)
        self.assertIn("Resumen datos", wb.sheetnames)
        self.assertIn("BOM consolidado", wb.sheetnames)
        self.assertIn("Reglas calculo", wb.sheetnames)
        panel = wb["Resumen"]
        self.assertEqual(panel["A1"].value, "Pollyana's Dolce")
        self.assertEqual(panel["D1"].value, "BOM requerido de insumos y materia prima")
        self.assertEqual(panel["A5"].value, 60)
        ws = wb["BOM consolidado"]
        self.assertEqual(ws["A2"].value, "AZ-XLSX-01")
        self.assertEqual(ws["E2"].value, "Azúcar Cálculo XLSX")
        self.assertEqual(ws["F2"].value, 2)
        self.assertEqual(ws["G2"].value, "kg")
        self.assertEqual(ws["I2"].value, 60)

    def test_calculo_insumos_export_bloquea_observaciones_criticas(self):
        session = self.client.session
        session["calculo_insumos_preview"] = {
            "target_start": "2026-04-01",
            "target_end": "2026-04-01",
            "target_label": "Cálculo bloqueado",
            "alcance": "calculo_insumos",
            "periodo": "2026-04",
            "sucursal_nombre": "Producción",
            "source_label": "Prueba",
            "source_rows": [],
            "unresolved_rows": [{"codigo_point": "SIN-CODIGO", "producto": "Producto sin match", "motivo": "Sin producto final equivalente"}],
            "skipped_rows": [],
            "totals": {"forecast_total": 0.0, "recetas_count": 0},
            "rows": [],
        }
        session.save()

        response = self.client.get(reverse("recetas:calculo_insumos_export"))
        self.assertEqual(response.status_code, 302)
        self.assertNotEqual(
            response.get("Content-Type"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_comparativo_pronostico_vs_solicitud_escenario_bajo(self):
        for month_idx, qty in [(11, "60"), (12, "72"), (1, "81"), (2, "78"), (3, "90")]:
            year = 2025 if month_idx >= 11 else 2026
            VentaHistorica.objects.create(
                receta=self.receta,
                sucursal=self.sucursal,
                fecha=date(year, month_idx, 15),
                cantidad=Decimal(qty),
                fuente="TEST_HIST_SOL",
            )

        SolicitudVenta.objects.create(
            receta=self.receta,
            sucursal=self.sucursal,
            alcance=SolicitudVenta.ALCANCE_MES,
            periodo="2026-04",
            fecha_inicio=date(2026, 4, 1),
            fecha_fin=date(2026, 4, 30),
            cantidad=Decimal("110"),
            fuente="TEST_SOL",
        )

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "mes",
                "periodo": "2026-04",
                "sucursal_id": str(self.sucursal.id),
                "run_mode": "preview",
                "escenario": "base",
                "safety_pct": "0",
            },
        )
        self.assertEqual(response.status_code, 302)

        response = self.client.get(
            reverse("recetas:plan_produccion"),
            {"periodo": "2026-04", "forecast_compare_escenario": "bajo"},
        )
        self.assertEqual(response.status_code, 200)
        compare = response.context["forecast_vs_solicitud"]
        row = next((r for r in compare["rows"] if r["receta_id"] == self.receta.id), None)
        self.assertIsNotNone(row)
        self.assertEqual(compare["escenario"], "bajo")
        self.assertEqual(row["forecast_qty"], row["forecast_low"])

    def test_export_pronostico_vs_solicitud_csv_y_xlsx(self):
        for month_idx, qty in [(11, "60"), (12, "72"), (1, "81"), (2, "78"), (3, "90")]:
            year = 2025 if month_idx >= 11 else 2026
            VentaHistorica.objects.create(
                receta=self.receta,
                sucursal=self.sucursal,
                fecha=date(year, month_idx, 15),
                cantidad=Decimal(qty),
                fuente="TEST_HIST_SOL_EXPORT",
            )

        SolicitudVenta.objects.create(
            receta=self.receta,
            sucursal=self.sucursal,
            alcance=SolicitudVenta.ALCANCE_MES,
            periodo="2026-04",
            fecha_inicio=date(2026, 4, 1),
            fecha_fin=date(2026, 4, 30),
            cantidad=Decimal("110"),
            fuente="TEST_SOL_EXPORT",
        )

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "mes",
                "periodo": "2026-04",
                "sucursal_id": str(self.sucursal.id),
                "run_mode": "preview",
                "escenario": "base",
                "safety_pct": "0",
            },
        )
        self.assertEqual(response.status_code, 302)

        response_csv = self.client.get(
            reverse("recetas:forecast_vs_solicitud_export"),
            {"format": "csv", "escenario": "bajo", "periodo": "2026-04"},
        )
        self.assertEqual(response_csv.status_code, 200)
        self.assertIn("text/csv", response_csv["Content-Type"])
        body_csv = response_csv.content.decode("utf-8")
        self.assertIn("receta_id,receta,forecast,forecast_base,forecast_baja,forecast_alta", body_csv)
        self.assertIn("Pastel Solicitud", body_csv)

        response_xlsx = self.client.get(
            reverse("recetas:forecast_vs_solicitud_export"),
            {"format": "xlsx", "escenario": "alto", "periodo": "2026-04"},
        )
        self.assertEqual(response_xlsx.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response_xlsx["Content-Type"],
        )
        wb = load_workbook(BytesIO(response_xlsx.content), data_only=True)
        self.assertIn("Resumen", wb.sheetnames)
        self.assertIn("Detalle", wb.sheetnames)

    def test_aplicar_ajuste_desde_forecast_actualiza_solicitud(self):
        for month_idx, qty in [(11, "60"), (12, "72"), (1, "81"), (2, "78"), (3, "90")]:
            year = 2025 if month_idx >= 11 else 2026
            VentaHistorica.objects.create(
                receta=self.receta,
                sucursal=self.sucursal,
                fecha=date(year, month_idx, 15),
                cantidad=Decimal(qty),
                fuente="TEST_HIST_SOL",
            )

        SolicitudVenta.objects.create(
            receta=self.receta,
            sucursal=self.sucursal,
            alcance=SolicitudVenta.ALCANCE_MES,
            periodo="2026-04",
            fecha_inicio=date(2026, 4, 1),
            fecha_fin=date(2026, 4, 30),
            cantidad=Decimal("110"),
            fuente="TEST_SOL",
        )

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "mes",
                "periodo": "2026-04",
                "sucursal_id": str(self.sucursal.id),
                "run_mode": "preview",
                "safety_pct": "0",
            },
        )
        self.assertEqual(response.status_code, 302)

        response = self.client.get(reverse("recetas:plan_produccion"), {"periodo": "2026-04"})
        self.assertEqual(response.status_code, 200)
        compare = response.context["forecast_vs_solicitud"]
        row = next((r for r in compare["rows"] if r["receta_id"] == self.receta.id), None)
        self.assertIsNotNone(row)
        expected_qty = Decimal(str(row["forecast_qty"]))

        response = self.client.post(
            reverse("recetas:solicitud_ventas_aplicar_desde_forecast"),
            {
                "modo": "receta",
                "receta_id": str(self.receta.id),
            },
        )
        self.assertEqual(response.status_code, 302)

        solicitud = SolicitudVenta.objects.get(
            receta=self.receta,
            sucursal=self.sucursal,
            alcance=SolicitudVenta.ALCANCE_MES,
            fecha_inicio=date(2026, 4, 1),
            fecha_fin=date(2026, 4, 30),
        )
        self.assertEqual(solicitud.cantidad, expected_qty)

    def test_aplicar_ajuste_desde_forecast_con_tope_omite_cambios_grandes(self):
        for month_idx, qty in [(11, "60"), (12, "72"), (1, "81"), (2, "78"), (3, "90")]:
            year = 2025 if month_idx >= 11 else 2026
            VentaHistorica.objects.create(
                receta=self.receta,
                sucursal=self.sucursal,
                fecha=date(year, month_idx, 15),
                cantidad=Decimal(qty),
                fuente="TEST_HIST_CAP",
            )

        SolicitudVenta.objects.create(
            receta=self.receta,
            sucursal=self.sucursal,
            alcance=SolicitudVenta.ALCANCE_MES,
            periodo="2026-04",
            fecha_inicio=date(2026, 4, 1),
            fecha_fin=date(2026, 4, 30),
            cantidad=Decimal("500"),
            fuente="TEST_SOL_CAP",
        )

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "mes",
                "periodo": "2026-04",
                "sucursal_id": str(self.sucursal.id),
                "run_mode": "preview",
                "safety_pct": "0",
            },
        )
        self.assertEqual(response.status_code, 302)

        response = self.client.post(
            reverse("recetas:solicitud_ventas_aplicar_desde_forecast"),
            {
                "modo": "receta",
                "receta_id": str(self.receta.id),
                "max_variacion_pct": "10",
            },
        )
        self.assertEqual(response.status_code, 302)

        solicitud = SolicitudVenta.objects.get(
            receta=self.receta,
            sucursal=self.sucursal,
            alcance=SolicitudVenta.ALCANCE_MES,
            fecha_inicio=date(2026, 4, 1),
            fecha_fin=date(2026, 4, 30),
        )
        self.assertEqual(solicitud.cantidad, Decimal("500"))

    def test_aplicar_ajuste_desde_forecast_usa_escenario_bajo(self):
        for month_idx, qty in [(11, "60"), (12, "72"), (1, "81"), (2, "78"), (3, "90")]:
            year = 2025 if month_idx >= 11 else 2026
            VentaHistorica.objects.create(
                receta=self.receta,
                sucursal=self.sucursal,
                fecha=date(year, month_idx, 15),
                cantidad=Decimal(qty),
                fuente="TEST_HIST_ESC",
            )

        SolicitudVenta.objects.create(
            receta=self.receta,
            sucursal=self.sucursal,
            alcance=SolicitudVenta.ALCANCE_MES,
            periodo="2026-04",
            fecha_inicio=date(2026, 4, 1),
            fecha_fin=date(2026, 4, 30),
            cantidad=Decimal("110"),
            fuente="TEST_SOL_ESC",
        )

        response = self.client.post(
            reverse("recetas:pronostico_estadistico_desde_historial"),
            {
                "alcance": "mes",
                "periodo": "2026-04",
                "sucursal_id": str(self.sucursal.id),
                "run_mode": "preview",
                "escenario": "bajo",
                "safety_pct": "0",
            },
        )
        self.assertEqual(response.status_code, 302)

        response = self.client.get(reverse("recetas:plan_produccion"), {"periodo": "2026-04"})
        self.assertEqual(response.status_code, 200)
        compare = response.context["forecast_vs_solicitud"]
        row = next((r for r in compare["rows"] if r["receta_id"] == self.receta.id), None)
        self.assertIsNotNone(row)
        expected_qty = Decimal(str(row["forecast_low"]))

        response = self.client.post(
            reverse("recetas:solicitud_ventas_aplicar_desde_forecast"),
            {
                "modo": "receta",
                "receta_id": str(self.receta.id),
                "escenario": "bajo",
            },
        )
        self.assertEqual(response.status_code, 302)

        solicitud = SolicitudVenta.objects.get(
            receta=self.receta,
            sucursal=self.sucursal,
            alcance=SolicitudVenta.ALCANCE_MES,
            fecha_inicio=date(2026, 4, 1),
            fecha_fin=date(2026, 4, 30),
        )
        self.assertEqual(solicitud.cantidad, expected_qty)


class RecetaPhase2ViewsTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_phase2",
            email="admin_phase2@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

        self.unidad = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.insumo = Insumo.objects.create(nombre="Insumo Driver", unidad_base=self.unidad, activo=True)
        self.receta = Receta.objects.create(
            nombre="Receta Driver",
            sheet_name="Insumos 1",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("8"),
            rendimiento_unidad=self.unidad,
            hash_contenido="hash-phase2-views-001",
        )
        self.linea = LineaReceta.objects.create(
            receta=self.receta,
            posicion=1,
            insumo=self.insumo,
            insumo_texto="Insumo Driver",
            cantidad=Decimal("2"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("4"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        asegurar_version_costeo(self.receta, fuente="TEST_PHASE2")
        self.linea.costo_unitario_snapshot = Decimal("5")
        self.linea.save(update_fields=["costo_unitario_snapshot"])
        asegurar_version_costeo(self.receta, fuente="TEST_PHASE2")

    def test_receta_detail_renderiza_comparador_y_export(self):
        resp = self.client.get(reverse("recetas:receta_detail", args=[self.receta.id]))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Comparar versiones")
        self.assertContains(resp, "Exportar historial CSV")
        self.assertContains(resp, "Parámetros de costeo")
        self.assertContains(resp, "Familia comercial (seleccionar)")
        self.assertIn("Pastel", resp.context["familias_catalogo"])

    def test_receta_detail_shows_supply_chain_for_base_recipe(self):
        self.receta.usa_presentaciones = True
        self.receta.save(update_fields=["usa_presentaciones"])
        derivado = Insumo.objects.create(
            nombre="Pan Derivado QA",
            codigo=f"DERIVADO:RECETA:{self.receta.id}:PRESENTACION:CHICO",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )
        producto_final = Receta.objects.create(
            nombre="Pastel QA Cadena",
            hash_contenido="hash-phase2-supply-chain-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=producto_final,
            posicion=1,
            insumo=derivado,
            insumo_texto=derivado.nombre,
            cantidad=Decimal("2"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("7"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        resp = self.client.get(reverse("recetas:receta_detail", args=[self.receta.id]))

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Cadena ERP")
        self.assertContains(resp, "Productos finales que lo consumen")
        self.assertContains(resp, "Pastel QA Cadena")

    def test_producto_final_detail_shows_source_base_for_derived_component(self):
        base = Receta.objects.create(
            nombre="Base QA Origen",
            hash_contenido="hash-phase2-source-link-001",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("5"),
            rendimiento_unidad=self.unidad,
        )
        derivado = Insumo.objects.create(
            nombre="Derivado QA Origen",
            codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:MINI",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )
        producto_final = Receta.objects.create(
            nombre="Pastel QA Origen",
            hash_contenido="hash-phase2-source-link-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=producto_final,
            posicion=1,
            insumo=derivado,
            insumo_texto=derivado.nombre,
            cantidad=Decimal("1"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("8"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        resp = self.client.get(reverse("recetas:receta_detail", args=[producto_final.id]))

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Base origen:")
        self.assertContains(resp, "Base QA Origen")

    def test_producto_final_detail_shows_operational_dependency_snapshot(self):
        base = Receta.objects.create(
            nombre="Base QA Dependencia",
            hash_contenido="hash-phase2-dependency-001",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("4"),
            rendimiento_unidad=self.unidad,
        )
        derivado = Insumo.objects.create(
            nombre="Derivado QA Dependencia",
            codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:MEDIANO",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )
        producto_final = Receta.objects.create(
            nombre="Pastel QA Dependencia",
            hash_contenido="hash-phase2-dependency-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=producto_final,
            posicion=1,
            insumo=derivado,
            insumo_texto=derivado.nombre,
            cantidad=Decimal("1.5"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("9"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        resp = self.client.get(reverse("recetas:receta_detail", args=[producto_final.id]))

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Dependencia operativa")
        self.assertContains(resp, "Bases internas origen")
        self.assertContains(resp, "Base QA Dependencia")

    def test_producto_final_detail_warns_when_internal_has_no_base_origin(self):
        interno = Insumo.objects.create(
            nombre="Interno sin origen QA",
            codigo="INT-QA-SIN-ORIGEN",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )
        producto_final = Receta.objects.create(
            nombre="Pastel QA Sin Origen",
            hash_contenido="hash-phase2-no-origin-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=producto_final,
            posicion=1,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("6"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        resp = self.client.get(reverse("recetas:receta_detail", args=[producto_final.id]))

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Dependencia operativa")
        self.assertContains(resp, "sin trazabilidad de base")

    def test_producto_final_detail_warns_when_using_base_direct_with_active_presentaciones(self):
        base = Receta.objects.create(
            nombre="Base QA Directa",
            hash_contenido="hash-phase2-direct-base-001",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("6"),
            rendimiento_unidad=self.unidad,
        )
        RecetaPresentacion.objects.create(
            receta=base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.380000"),
            activo=True,
        )
        base_directa = Insumo.objects.create(
            nombre="Base QA Directa",
            codigo=f"DERIVADO:RECETA:{base.id}:PREPARACION",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )
        producto_final = Receta.objects.create(
            nombre="Pastel QA Base Directa",
            hash_contenido="hash-phase2-direct-base-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=producto_final,
            posicion=1,
            insumo=base_directa,
            insumo_texto=base_directa.nombre,
            cantidad=Decimal("0.380000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("11"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        resp = self.client.get(reverse("recetas:receta_detail", args=[producto_final.id]))

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Producto final usando base sin presentación")
        self.assertContains(resp, "Base QA Directa")
        self.assertContains(resp, "Usa base sin presentación")
        self.assertContains(resp, "Base QA Directa - Chico")
        self.assertContains(resp, "Coincide con la cantidad capturada")

    def test_producto_final_detail_ignores_implausible_direct_base_replacement(self):
        base = Receta.objects.create(
            nombre="Base QA Directa Implausible",
            hash_contenido="hash-phase2-direct-base-implausible-001",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("6"),
            rendimiento_unidad=self.unidad,
        )
        presentacion = RecetaPresentacion.objects.create(
            receta=base,
            nombre="Individual",
            peso_por_unidad_kg=Decimal("0.090000"),
            activo=True,
        )
        Insumo.objects.create(
            nombre="Base QA Directa Implausible - Individual",
            codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:{presentacion.id}",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )
        base_directa = Insumo.objects.create(
            nombre="Base QA Directa Implausible",
            codigo=f"DERIVADO:RECETA:{base.id}:PREPARACION",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )
        empaque = Insumo.objects.create(
            nombre="Empaque QA Implausible",
            codigo="EMP-QA-IMPLAUSIBLE",
            tipo_item=Insumo.TIPO_EMPAQUE,
            unidad_base=self.unidad,
            activo=True,
            categoria="Empaques",
        )
        producto_final = Receta.objects.create(
            nombre="Pastel QA Base Directa Implausible",
            hash_contenido="hash-phase2-direct-base-implausible-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=producto_final,
            posicion=1,
            insumo=base_directa,
            insumo_texto=base_directa.nombre,
            cantidad=Decimal("8.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("11"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        LineaReceta.objects.create(
            receta=producto_final,
            posicion=2,
            insumo=empaque,
            insumo_texto=empaque.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="pza",
            costo_unitario_snapshot=Decimal("1"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        resp = self.client.get(reverse("recetas:receta_detail", args=[producto_final.id]))

        self.assertEqual(resp.status_code, 200)
        self.assertNotContains(resp, "Producto final usando base sin presentación")
        self.assertNotContains(resp, "Usa base sin presentación")

    def test_producto_final_can_apply_suggested_direct_base_replacement(self):
        base = Receta.objects.create(
            nombre="Base QA Reemplazo",
            hash_contenido="hash-phase2-direct-base-replace-001",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("6"),
            rendimiento_unidad=self.unidad,
        )
        presentacion = RecetaPresentacion.objects.create(
            receta=base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.380000"),
            activo=True,
        )
        derivado = Insumo.objects.create(
            nombre="Base QA Reemplazo - Chico",
            codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:{presentacion.id}",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )
        base_directa = Insumo.objects.create(
            nombre="Base QA Reemplazo",
            codigo=f"DERIVADO:RECETA:{base.id}:PREPARACION",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )
        producto_final = Receta.objects.create(
            nombre="Pastel QA Reemplazo Directo",
            hash_contenido="hash-phase2-direct-base-replace-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        linea = LineaReceta.objects.create(
            receta=producto_final,
            posicion=1,
            insumo=base_directa,
            insumo_texto=base_directa.nombre,
            cantidad=Decimal("0.380000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("11"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.post(
            reverse("recetas:linea_apply_direct_base_replacement", args=[producto_final.id, linea.id]),
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        linea.refresh_from_db()
        self.assertEqual(linea.insumo_id, derivado.id)
        self.assertEqual(linea.insumo_texto, derivado.nombre)
        self.assertEqual(linea.cantidad, Decimal("1.000000"))
        self.assertContains(response, "Línea actualizada a Base QA Reemplazo - Chico")

    def test_producto_final_can_apply_suggested_direct_base_replacements_in_batch(self):
        base = Receta.objects.create(
            nombre="Base QA Lote",
            hash_contenido="hash-phase2-direct-base-batch-001",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("8"),
            rendimiento_unidad=self.unidad,
        )
        presentacion = RecetaPresentacion.objects.create(
            receta=base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.400000"),
            activo=True,
        )
        derivado = Insumo.objects.create(
            nombre="Base QA Lote - Chico",
            codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:{presentacion.id}",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )
        base_directa = Insumo.objects.create(
            nombre="Base QA Lote",
            codigo=f"DERIVADO:RECETA:{base.id}:PREPARACION",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )
        producto_final = Receta.objects.create(
            nombre="Pastel QA Lote Directo",
            hash_contenido="hash-phase2-direct-base-batch-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        linea_1 = LineaReceta.objects.create(
            receta=producto_final,
            posicion=1,
            insumo=base_directa,
            insumo_texto=base_directa.nombre,
            cantidad=Decimal("0.400000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("11"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        linea_2 = LineaReceta.objects.create(
            receta=producto_final,
            posicion=2,
            insumo=base_directa,
            insumo_texto=base_directa.nombre,
            cantidad=Decimal("0.800000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("11"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.post(
            reverse("recetas:receta_apply_direct_base_replacements", args=[producto_final.id]),
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        linea_1.refresh_from_db()
        linea_2.refresh_from_db()
        self.assertEqual(linea_1.insumo_id, derivado.id)
        self.assertEqual(linea_2.insumo_id, derivado.id)
        self.assertEqual(linea_1.cantidad, Decimal("1.000000"))
        self.assertEqual(linea_2.cantidad, Decimal("2.000000"))
        self.assertContains(response, "Se actualizaron 2 línea(s) a su presentación derivada sugerida.")

    def test_producto_final_detail_shows_incomplete_erp_item_warning(self):
        interno = Insumo.objects.create(
            nombre="Interno ERP Incompleto QA",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            categoria="",
            activo=True,
        )
        producto_final = Receta.objects.create(
            nombre="Pastel QA ERP Incompleto",
            hash_contenido="hash-phase2-erp-incomplete-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=producto_final,
            posicion=1,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("6"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        sucursal = Sucursal.objects.create(codigo="SUC-ERP-CRIT", nombre="Sucursal ERP Critica", activa=True)
        VentaHistorica.objects.create(
            receta=producto_final,
            sucursal=sucursal,
            fecha=timezone.localdate() - timedelta(days=2),
            cantidad=Decimal("85"),
            tickets=8,
        )

        resp = self.client.get(reverse("recetas:receta_detail", args=[producto_final.id]))

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Demanda crítica bloqueada por maestro")
        self.assertContains(resp, "Demanda crítica bloqueada")
        self.assertContains(resp, "Falta: categoría")

    def test_receta_update_producto_final_requires_familia(self):
        payload = {
            "nombre": "Receta Driver",
            "codigo_point": "",
            "familia": "",
            "categoria": "",
            "sheet_name": "Insumos 1",
            "tipo": Receta.TIPO_PRODUCTO_FINAL,
            "rendimiento_cantidad": "8",
            "rendimiento_unidad_id": str(self.unidad.id),
        }
        resp = self.client.post(reverse("recetas:receta_update", args=[self.receta.id]), payload)
        self.assertEqual(resp.status_code, 302)
        self.receta.refresh_from_db()
        self.assertEqual(self.receta.tipo, Receta.TIPO_PREPARACION)

    def test_receta_update_base_requires_rendimiento(self):
        payload = {
            "nombre": "Receta Driver",
            "codigo_point": "",
            "familia": "",
            "categoria": "",
            "sheet_name": "Insumos 1",
            "tipo": Receta.TIPO_PREPARACION,
            "rendimiento_cantidad": "",
            "rendimiento_unidad_id": "",
        }
        resp = self.client.post(reverse("recetas:receta_update", args=[self.receta.id]), payload, follow=True)
        self.assertEqual(resp.status_code, 200)
        self.receta.refresh_from_db()
        self.assertEqual(self.receta.rendimiento_cantidad, Decimal("8"))
        self.assertContains(resp, "Debes capturar el rendimiento total de la batida para costeo enterprise.")

    def test_receta_update_rejects_switch_to_base_without_clearing_presentaciones(self):
        receta = Receta.objects.create(
            nombre="Receta Presentaciones QA",
            sheet_name="Insumos 1",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("10"),
            rendimiento_unidad=self.unidad,
            hash_contenido="hash-phase2-views-presentaciones-001",
        )
        RecetaPresentacion.objects.create(
            receta=receta,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.380000"),
            activo=True,
        )
        payload = {
            "nombre": receta.nombre,
            "codigo_point": "",
            "familia": "",
            "categoria": "",
            "sheet_name": "Insumos 1",
            "tipo": Receta.TIPO_PREPARACION,
            "rendimiento_cantidad": "10",
            "rendimiento_unidad_id": str(self.unidad.id),
        }
        resp = self.client.post(reverse("recetas:receta_update", args=[receta.id]), payload, follow=True)
        self.assertEqual(resp.status_code, 200)
        receta.refresh_from_db()
        self.assertTrue(receta.usa_presentaciones)
        self.assertContains(resp, "No puedes convertir a base simple una receta que ya tiene presentaciones activas.")

    def test_receta_versiones_export_csv(self):
        resp = self.client.get(
            reverse("recetas:receta_versiones_export", args=[self.receta.id]),
            {"format": "csv"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertIn("text/csv", resp["Content-Type"])
        self.assertIn("version,fecha,fuente", resp.content.decode("utf-8"))

    def test_receta_versiones_export_xlsx(self):
        resp = self.client.get(
            reverse("recetas:receta_versiones_export", args=[self.receta.id]),
            {"format": "xlsx"},
        )
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheetml", resp["Content-Type"])

    def test_receta_detail_handles_missing_version_table_gracefully(self):
        with patch("recetas.views._load_versiones_costeo", side_effect=OperationalError("missing table")):
            resp = self.client.get(reverse("recetas:receta_detail", args=[self.receta.id]))
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.context["versiones_unavailable"])
        self.assertContains(resp, "no está disponible en este entorno")

    def test_receta_detail_handles_missing_driver_table_gracefully(self):
        with patch("recetas.views.calcular_costeo_receta", side_effect=OperationalError("missing table")):
            resp = self.client.get(reverse("recetas:receta_detail", args=[self.receta.id]))
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.context["costeo_unavailable"])
        self.assertContains(resp, "costeo avanzado por parámetros no está disponible")

    def test_receta_versiones_export_handles_missing_version_table_gracefully(self):
        with patch("recetas.views._load_versiones_costeo", side_effect=OperationalError("missing table")):
            resp = self.client.get(
                reverse("recetas:receta_versiones_export", args=[self.receta.id]),
                {"format": "csv"},
            )
        self.assertEqual(resp.status_code, 200)
        self.assertIn("text/csv", resp["Content-Type"])
        self.assertIn("version,fecha,fuente", resp.content.decode("utf-8"))

    def test_drivers_costeo_create_and_list(self):
        payload = {
            "scope": CostoDriver.SCOPE_PRODUCTO,
            "nombre": "Driver test producto",
            "receta_id": str(self.receta.id),
            "familia": "",
            "lote_desde": "",
            "lote_hasta": "",
            "mo_pct": "7",
            "indirecto_pct": "3",
            "mo_fijo": "0",
            "indirecto_fijo": "0",
            "prioridad": "15",
            "activo": "1",
        }
        post_resp = self.client.post(reverse("recetas:drivers_costeo"), payload)
        self.assertEqual(post_resp.status_code, 302)
        self.assertEqual(CostoDriver.objects.count(), 1)
        get_resp = self.client.get(reverse("recetas:drivers_costeo"))
        self.assertEqual(get_resp.status_code, 200)
        self.assertContains(get_resp, "Resumen de drivers")
        self.assertContains(get_resp, "Flujo del costeo")
        self.assertContains(get_resp, "Driver test producto")

    def test_drivers_costeo_plantilla_csv(self):
        resp = self.client.get(reverse("recetas:drivers_costeo_plantilla"), {"format": "csv"})
        self.assertEqual(resp.status_code, 200)
        self.assertIn("text/csv", resp["Content-Type"])
        body = resp.content.decode("utf-8")
        self.assertIn("scope,nombre,receta", body)

    def test_drivers_costeo_plantilla_xlsx_valida(self):
        resp = self.client.get(reverse("recetas:drivers_costeo_plantilla"), {"format": "xlsx"})
        self.assertEqual(resp.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resp["Content-Type"],
        )
        wb = load_workbook(filename=BytesIO(resp.content), data_only=True)
        self.assertIn("drivers_costeo", wb.sheetnames)
        self.assertIn("instrucciones", wb.sheetnames)
        ws = wb["drivers_costeo"]
        self.assertEqual(ws["A1"].value, "scope")
        self.assertEqual(ws["B1"].value, "nombre")
        self.assertEqual(ws["A2"].value, "PRODUCTO")

    def test_drivers_costeo_handles_missing_table_gracefully(self):
        with patch("recetas.views.CostoDriver.objects.select_related", side_effect=OperationalError("missing table")):
            resp = self.client.get(reverse("recetas:drivers_costeo"))
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.context["drivers_unavailable"])
        self.assertContains(resp, "no disponibles en este entorno")


class ImportCosteoDriverExcelTests(TestCase):
    def setUp(self):
        self.tempfile = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb = Workbook()

        ws_cost = wb.active
        ws_cost.title = "Costo Materia Prima"
        ws_cost.append(["Proveedor", "Producto", "Descripcion", "Cantidad", "Unidad", "Costo", "Fecha"])
        ws_cost.append(["Proveedor Test", "Harina Driver", "", 1, "kg", 10, "2026-02-20"])

        ws_receta = wb.create_sheet("Insumos 1")
        ws_receta.append(["Batida Driver"])
        ws_receta.append(["Ingrediente", "Cantidad", "Unidad", "Costo"])
        ws_receta.append(["Harina Driver", 1, "kg", 10])

        ws_drivers = wb.create_sheet("Drivers Costeo")
        ws_drivers.append(
            [
                "scope",
                "nombre",
                "receta",
                "mo_pct",
                "indirecto_pct",
                "mo_fijo",
                "indirecto_fijo",
                "prioridad",
                "activo",
            ]
        )
        ws_drivers.append(
            [
                "PRODUCTO",
                "Driver Batida Import",
                "Batida Driver",
                10,
                5,
                0,
                0,
                10,
                1,
            ]
        )

        wb.save(self.tempfile.name)
        wb.close()

    def tearDown(self):
        try:
            os.unlink(self.tempfile.name)
        except OSError:
            pass

    def test_import_costeo_detecta_drivers_y_versiona_costeo(self):
        resultado = ImportadorCosteo(self.tempfile.name).procesar_completo()
        self.assertEqual(resultado.drivers_hojas_detectadas, 1)
        self.assertEqual(resultado.drivers_creados, 1)

        receta = Receta.objects.get(nombre_normalizado="batida driver")
        driver = CostoDriver.objects.get(nombre="Driver Batida Import")
        self.assertEqual(driver.scope, CostoDriver.SCOPE_PRODUCTO)
        self.assertEqual(driver.receta_id, receta.id)

        latest = receta.versiones_costo.order_by("-version_num").first()
        self.assertIsNotNone(latest)
        self.assertGreater(latest.costo_mo, Decimal("0"))
        self.assertGreater(latest.costo_indirecto, Decimal("0"))


class ReabastoCedisSecurityAndFolioTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        ventas_group, _ = Group.objects.get_or_create(name=ROLE_VENTAS)
        compras_group, _ = Group.objects.get_or_create(name=ROLE_COMPRAS)

        self.user_ventas_no_sucursal = user_model.objects.create_user(
            username="ventas_sin_sucursal",
            email="ventas_sin_sucursal@example.com",
            password="test12345",
        )
        self.user_ventas_no_sucursal.groups.add(ventas_group)

        self.user_ventas_sucursal = user_model.objects.create_user(
            username="ventas_sucursal",
            email="ventas_sucursal@example.com",
            password="test12345",
        )
        self.user_ventas_sucursal.groups.add(ventas_group)
        self.user_branch_capture = user_model.objects.create_user(
            username="sucursal_captura",
            email="sucursal_captura@example.com",
            password="test12345",
        )
        self.user_branch_capture.groups.add(ventas_group)

        self.user_compras = user_model.objects.create_user(
            username="compras_reabasto",
            email="compras_reabasto@example.com",
            password="test12345",
        )
        self.user_compras.groups.add(compras_group)

        self.sucursal_colosio = Sucursal.objects.create(codigo="COLOSIO", nombre="Colosio", activa=True)
        self.sucursal_leyva = Sucursal.objects.create(codigo="LEYVA", nombre="Leyva", activa=True)
        UserProfile.objects.create(user=self.user_ventas_sucursal, sucursal=self.sucursal_leyva)
        UserProfile.objects.create(
            user=self.user_branch_capture,
            sucursal=self.sucursal_leyva,
            modo_captura_sucursal=True,
        )

        self.receta = Receta.objects.create(
            nombre="Pastel Seguridad Reabasto",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-reabasto-security-001",
        )

    def test_no_sucursal_profile_user_cannot_capture_reabasto(self):
        self.client.force_login(self.user_ventas_no_sucursal)
        before = SolicitudReabastoCedis.objects.count()
        response = self.client.post(
            reverse("recetas:reabasto_cedis_linea_guardar"),
            {
                "fecha_operacion": "2026-02-26",
                "sucursal_id": self.sucursal_colosio.id,
                "receta_id": self.receta.id,
                "stock_reportado": "1",
            },
        )
        self.assertEqual(response.status_code, 403)
        self.assertEqual(SolicitudReabastoCedis.objects.count(), before)

    def test_branch_user_cannot_capture_other_branch(self):
        self.client.force_login(self.user_ventas_sucursal)
        response = self.client.post(
            reverse("recetas:reabasto_cedis_linea_guardar"),
            {
                "fecha_operacion": "2026-02-26",
                "sucursal_id": self.sucursal_colosio.id,
                "receta_id": self.receta.id,
                "stock_reportado": "1",
            },
        )
        self.assertEqual(response.status_code, 403)

    def test_branch_user_can_capture_assigned_branch(self):
        self.client.force_login(self.user_ventas_sucursal)
        response = self.client.post(
            reverse("recetas:reabasto_cedis_linea_guardar"),
            {
                "fecha_operacion": "2026-02-26",
                "sucursal_id": self.sucursal_leyva.id,
                "receta_id": self.receta.id,
                "stock_reportado": "2",
                "en_transito": "1",
                "consumo_proyectado": "0.5",
            },
        )
        self.assertEqual(response.status_code, 302)
        solicitud = SolicitudReabastoCedis.objects.get(
            fecha_operacion=date(2026, 2, 26),
            sucursal=self.sucursal_leyva,
        )
        self.assertTrue(
            SolicitudReabastoCedisLinea.objects.filter(
                solicitud=solicitud,
                receta=self.receta,
            ).exists()
        )

    def test_branch_capture_mode_redirects_to_dedicated_capture_view(self):
        self.client.force_login(self.user_branch_capture)
        response = self.client.get(reverse("recetas:reabasto_cedis"))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("recetas:reabasto_cedis_captura"), response.url)

    def test_branch_capture_view_renders_for_assigned_branch(self):
        self.client.force_login(self.user_branch_capture)
        response = self.client.get(reverse("recetas:reabasto_cedis_captura"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Captura de cierre")
        self.assertContains(response, "Resumen del cálculo")
        self.assertContains(response, "Expediente ERP del cierre sucursal")
        self.assertContains(response, "Workflow ERP del cierre")
        self.assertContains(response, "Radar ejecutivo ERP")
        self.assertContains(response, "Mesa de gobierno ERP")
        self.assertContains(response, "<th>Responsable</th>", html=True)
        self.assertContains(response, "<th>Cierre</th>", html=True)
        self.assertContains(response, "<th>Siguiente paso</th>", html=True)
        self.assertIn("erp_command_center", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertIn("erp_governance_rows", response.context)
        self.assertContains(response, "LEYVA")

    def test_cierre_guardar_redirects_back_to_capture_view_for_branch_mode(self):
        PoliticaStockSucursalProducto.objects.create(
            sucursal=self.sucursal_leyva,
            receta=self.receta,
            stock_minimo=Decimal("5"),
            stock_objetivo=Decimal("5"),
            stock_maximo=Decimal("8"),
            dias_cobertura=1,
            stock_seguridad=Decimal("0"),
            lote_minimo=Decimal("0"),
            multiplo_empaque=Decimal("1"),
            activa=True,
        )
        self.client.force_login(self.user_branch_capture)
        response = self.client.post(
            reverse("recetas:reabasto_cedis_cierre_guardar"),
            {
                "fecha_operacion": "2026-02-26",
                "sucursal_id": self.sucursal_leyva.id,
                "row_receta_id": [str(self.receta.id)],
                f"stock_reportado_{self.receta.id}": "1",
                "accion": "BORRADOR",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("recetas:reabasto_cedis_captura"), response.url)

    def test_compras_user_can_update_estado_without_sucursal_profile(self):
        solicitud = SolicitudReabastoCedis.objects.create(
            fecha_operacion=date(2026, 2, 26),
            sucursal=self.sucursal_colosio,
        )
        self.client.force_login(self.user_compras)
        response = self.client.post(
            reverse("recetas:reabasto_cedis_estado_guardar", kwargs={"solicitud_id": solicitud.id}),
            {"estado": SolicitudReabastoCedis.ESTADO_ATENDIDA},
        )
        self.assertEqual(response.status_code, 302)
        solicitud.refresh_from_db()
        self.assertEqual(solicitud.estado, SolicitudReabastoCedis.ESTADO_ATENDIDA)

    def test_reabasto_folio_retries_on_unique_collision(self):
        SolicitudReabastoCedis.objects.create(
            folio="SRC-COLLIDE-001",
            fecha_operacion=date(2026, 2, 26),
            sucursal=self.sucursal_colosio,
        )
        with patch.object(
            SolicitudReabastoCedis,
            "_next_folio",
            side_effect=["SRC-COLLIDE-001", "SRC-COLLIDE-002"],
        ):
            solicitud = SolicitudReabastoCedis.objects.create(
                fecha_operacion=date(2026, 2, 26),
                sucursal=self.sucursal_leyva,
            )
        self.assertEqual(solicitud.folio, "SRC-COLLIDE-002")


class ReabastoCedisEnterpriseBoardTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_reabasto_board",
            email="admin_reabasto_board@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)
        self.unidad_kg = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.sucursal = Sucursal.objects.create(codigo="MATRIZ", nombre="Matriz", activa=True)
        self.receta_pendiente = Receta.objects.create(
            nombre="Pastel Reabasto Pendiente",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-reabasto-board-001",
        )
        self.receta_sin_empaque = Receta.objects.create(
            nombre="Pastel Reabasto Sin Empaque",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-reabasto-board-002",
            familia="Pasteles",
        )
        self.receta_sin_inventario = Receta.objects.create(
            nombre="Pastel Reabasto Sin Inventario",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-reabasto-board-003",
            familia="Pasteles",
        )
        self.insumo_interno = Insumo.objects.create(
            nombre="Betun Reabasto Interno",
            tipo_item=Insumo.TIPO_INTERNO,
            categoria="Betunes",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        self.materia_prima = Insumo.objects.create(
            nombre="Chocolate Reabasto MP",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            categoria="Cobertura",
            unidad_base=self.unidad_kg,
            activo=True,
            proveedor_principal=Proveedor.objects.create(nombre="Proveedor Reabasto MP"),
        )
        self.empaque = Insumo.objects.create(
            nombre="Caja Reabasto",
            tipo_item=Insumo.TIPO_EMPAQUE,
            categoria="Empaque",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        LineaReceta.objects.create(
            receta=self.receta_sin_empaque,
            posicion=1,
            insumo=self.materia_prima,
            insumo_texto=self.materia_prima.nombre,
            cantidad=Decimal("1"),
            unidad=self.unidad_kg,
            unidad_texto="kg",
            match_status=LineaReceta.STATUS_AUTO,
        )
        LineaReceta.objects.create(
            receta=self.receta_sin_inventario,
            posicion=1,
            insumo=self.insumo_interno,
            insumo_texto=self.insumo_interno.nombre,
            cantidad=Decimal("1"),
            unidad=self.unidad_kg,
            unidad_texto="kg",
            match_status=LineaReceta.STATUS_AUTO,
        )
        LineaReceta.objects.create(
            receta=self.receta_sin_inventario,
            posicion=2,
            insumo=self.empaque,
            insumo_texto=self.empaque.nombre,
            cantidad=Decimal("1"),
            unidad=self.unidad_kg,
            unidad_texto="kg",
            match_status=LineaReceta.STATUS_AUTO,
        )
        VentaHistorica.objects.create(
            receta=self.receta_sin_empaque,
            sucursal=self.sucursal,
            fecha=timezone.localdate() - timedelta(days=7),
            cantidad=Decimal("6"),
            monto_total=Decimal("600"),
            fuente="POINT_HIST_2026_Q1",
        )
        VentaHistorica.objects.create(
            receta=self.receta_sin_inventario,
            sucursal=self.sucursal,
            fecha=timezone.localdate() - timedelta(days=14),
            cantidad=Decimal("9"),
            monto_total=Decimal("900"),
            fuente="POINT_HIST_2026_Q1",
        )
        self.solicitud_reabasto = SolicitudReabastoCedis.objects.create(
            fecha_operacion=timezone.localdate(),
            sucursal=self.sucursal,
            creado_por=self.user,
        )
        SolicitudReabastoCedisLinea.objects.create(
            solicitud=self.solicitud_reabasto,
            receta=self.receta_sin_inventario,
            stock_reportado=Decimal("0"),
            en_transito=Decimal("0"),
            consumo_proyectado=Decimal("4"),
            sugerido=Decimal("4"),
            solicitado=Decimal("4"),
        )

    def test_reabasto_cedis_renders_enterprise_blockers(self):
        response = self.client.get(reverse("recetas:reabasto_cedis"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Resumen operativo CEDIS")
        self.assertContains(response, "Top decisiones del reabasto")
        self.assertContains(response, "Sucursales a priorizar hoy")
        self.assertContains(response, "Insumo a asegurar por sucursal")
        self.assertContains(response, "Cierre troncal ERP consolidado")
        self.assertContains(response, "Prioridades de atención")
        self.assertContains(response, "Dependencias del flujo")
        self.assertContains(response, "Sucursales / Plan")
        self.assertContains(response, "Compras documentales")
        self.assertContains(response, "Inventario / Reabasto")
        self.assertContains(response, "Resumen de seguimiento")
        self.assertContains(response, "Gate de generación")
        self.assertContains(response, "Base comparable")
        self.assertContains(response, "Años observados")
        self.assertContains(response, "Control de demanda comercial")
        self.assertContains(response, "Sucursales líderes")
        self.assertContains(response, "Productos líderes")
        self.assertContains(response, "Esperando cierres")
        self.assertIn("branch_priority_rows", response.context)
        self.assertTrue(response.context["branch_priority_rows"])
        self.assertIn("branch_supply_rows", response.context)
        self.assertTrue(response.context["branch_supply_rows"])
        self.assertContains(response, "Bloqueos enterprise para abastecimiento")
        self.assertContains(response, "Sin inventario CEDIS")
        self.assertContains(response, "Receta por validar")
        self.assertContains(response, "Sin empaque")
        self.assertContains(response, "Registrar inventario")
        self.assertContains(response, "Agregar empaque")
        self.assertContains(response, "Abrir cierres")
        self.assertContains(response, "Ver bloqueos")
        board = response.context["reabasto_enterprise_board"]
        demand_summary = response.context["demand_history_summary"]
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertTrue(demand_summary["available"])
        self.assertEqual(demand_summary["branch_count"], 1)
        self.assertEqual(demand_summary["recipe_count"], 2)
        self.assertIn("years_observed", demand_summary)
        self.assertIn("comparable_years", demand_summary)
        self.assertGreaterEqual(board["blocked_total"], 2)
        self.assertTrue(any(item["blocker_label"] == "Receta por validar" for item in board["detail_rows"]))
        self.assertTrue(any(item["blocker_label"] == "Sin empaque" for item in board["detail_rows"]))
        daily_control = response.context["reabasto_daily_control"]
        self.assertEqual(daily_control["stage_label"], "Esperando cierres")
        self.assertTrue(any(item["label"] == "Plan CEDIS" for item in daily_control["control_cards"]))
        self.assertFalse(daily_control["generation_gate"]["can_generate_plan"])
        self.assertFalse(daily_control["generation_gate"]["can_generate_compras"])
        self.assertIn("demand_gate", daily_control["generation_gate"])
        self.assertIn("trunk_handoff_rows", daily_control)
        self.assertEqual(len(daily_control["trunk_handoff_rows"]), 3)
        self.assertTrue(all(item.get("action_url") for item in daily_control["generation_gate"]["checks"]))
        self.assertIn("daily_decision_rows", response.context)
        self.assertTrue(response.context["daily_decision_rows"])
        self.assertIn("branch_priority_rows", response.context)
        self.assertTrue(response.context["branch_priority_rows"])
        compras_check = next(item for item in daily_control["generation_gate"]["checks"] if item["label"] == "Flujo de compras sin bloqueos")
        self.assertFalse(compras_check["is_ready"])
        self.assertIn("plan cedis", compras_check["detail"].lower())

    def test_reabasto_cedis_blocks_plan_generation_when_gate_fails(self):
        before = PlanProduccion.objects.count()
        with patch("recetas.views.log_event") as log_event_mock:
            response = self.client.post(
                reverse("recetas:reabasto_cedis_generar_plan"),
                {"fecha_operacion": timezone.localdate().isoformat()},
                follow=True,
            )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No se puede generar plan todavía")
        self.assertEqual(PlanProduccion.objects.count(), before)
        self.assertTrue(log_event_mock.called)
        self.assertEqual(log_event_mock.call_args.args[1], "BLOCKED")

    def test_reabasto_cedis_blocks_compras_generation_without_plan(self):
        before_solicitudes = SolicitudCompra.objects.count()
        before_ordenes = OrdenCompra.objects.count()
        with patch("recetas.views.log_event") as log_event_mock:
            response = self.client.post(
                reverse("recetas:reabasto_cedis_generar_compras"),
                {"fecha_operacion": timezone.localdate().isoformat()},
                follow=True,
            )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No se puede generar compras todavía")
        self.assertEqual(SolicitudCompra.objects.count(), before_solicitudes)
        self.assertEqual(OrdenCompra.objects.count(), before_ordenes)
        self.assertTrue(log_event_mock.called)
        self.assertEqual(log_event_mock.call_args.args[1], "BLOCKED")

    def test_reabasto_cedis_renders_document_pipeline_for_existing_plan(self):
        fecha_operacion = timezone.localdate()
        plan = PlanProduccion.objects.create(
            nombre=f"CEDIS Reabasto {fecha_operacion.isoformat()}",
            fecha_produccion=fecha_operacion,
            notas=f"[AUTO_REABASTO_CEDIS:{fecha_operacion.isoformat()}] prueba documental",
            creado_por=self.user,
        )
        PlanProduccionItem.objects.create(plan=plan, receta=self.receta_sin_empaque, cantidad=Decimal("2"))
        solicitud = SolicitudCompra.objects.create(
            area=f"PLAN_PRODUCCION:{plan.id}",
            solicitante="admin",
            insumo=self.materia_prima,
            proveedor_sugerido=self.materia_prima.proveedor_principal,
            cantidad=Decimal("5"),
            fecha_requerida=fecha_operacion,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        orden = OrdenCompra.objects.create(
            solicitud=solicitud,
            referencia=f"PLAN_PRODUCCION:{plan.id}",
            proveedor=self.materia_prima.proveedor_principal,
            fecha_emision=fecha_operacion,
            monto_estimado=Decimal("100"),
            estatus=OrdenCompra.STATUS_ENVIADA,
        )
        RecepcionCompra.objects.create(
            orden=orden,
            fecha_recepcion=fecha_operacion,
            conformidad_pct=Decimal("90"),
            estatus=RecepcionCompra.STATUS_PENDIENTE,
            observaciones="Recepcion abierta",
        )

        response = self.client.get(reverse("recetas:reabasto_cedis"), {"fecha": fecha_operacion.isoformat()})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cierre troncal ERP consolidado")
        self.assertContains(response, "Semáforo documental")
        self.assertContains(response, "Dependencias del flujo")
        self.assertContains(response, "Resumen de seguimiento")
        self.assertContains(response, "Control por frente")
        self.assertContains(response, "Bloqueo prioritario por etapa")
        self.assertContains(response, "Entregas entre etapas")
        self.assertContains(response, "Entrega prioritaria")
        self.assertContains(response, "Solicitudes")
        self.assertContains(response, "Órdenes")
        self.assertContains(response, "Recepciones")
        self.assertContains(response, "<th>Responsable</th>", html=True)
        self.assertContains(response, "<th>Cierre</th>", html=True)
        self.assertContains(response, "<th>Siguiente paso</th>", html=True)
        self.assertIn("executive_radar_rows", response.context)
        self.assertContains(response, "Cierre")
        self.assertContains(response, "Siguiente paso:")
        daily_control = response.context["reabasto_daily_control"]
        self.assertTrue(daily_control["document_stage_rows"])
        self.assertTrue(daily_control["pipeline_steps"])
        self.assertTrue(daily_control["document_blocker_rows"])
        self.assertIsNotNone(daily_control["purchase_gate"])
        self.assertIsNotNone(daily_control["stage_focus"])
        self.assertTrue(daily_control["handoff_checks"])
        self.assertIsNotNone(daily_control["handoff_focus"])
        self.assertTrue(daily_control["closure_focus_rows"])
        self.assertEqual(daily_control["closure_focus_rows"][0]["scope"], "Recepción")
        self.assertTrue(daily_control["handoff_focus_rows"])
        self.assertEqual(daily_control["handoff_focus_rows"][0]["scope"], "Recepción")
        self.assertEqual(daily_control["stage_focus"]["scope"], "Recepción")
        self.assertTrue(daily_control["stage_focus"]["blocker_rows"])
        self.assertTrue(all(row["scope"] == "Recepción" for row in daily_control["stage_focus"]["blocker_rows"]))
        self.assertEqual(daily_control["master_summary"]["label"], "Maestro ERP al día")
        self.assertEqual(daily_control["master_summary"]["blocked_count"], 0)
        self.assertEqual(daily_control["master_summary"]["progress_pct"], 100)
        self.assertEqual(len(daily_control["trunk_handoff_rows"]), 3)
        self.assertEqual(daily_control["document_stage_rows"][0]["owner"], "Compras / Solicitante")
        self.assertIn("next_step", daily_control["document_stage_rows"][0])
        self.assertIn("action_detail", daily_control["document_stage_rows"][0])
        self.assertEqual(daily_control["pipeline_steps"][0]["owner"], "Compras / Solicitante")
        self.assertIn("next_step", daily_control["pipeline_steps"][0])
        self.assertIn("action_detail", daily_control["pipeline_steps"][0])
        self.assertTrue(any(row["scope"] in {"Solicitud", "Orden", "Recepción"} for row in daily_control["document_blocker_rows"]))
        self.assertContains(response, "<th>Ámbito</th>", html=True)
        self.assertContains(response, "Maestro ERP al día")

    def test_reabasto_cedis_can_focus_master_blocker_class(self):
        fecha_operacion = timezone.localdate()
        plan = PlanProduccion.objects.create(
            nombre=f"CEDIS Reabasto {fecha_operacion.isoformat()}",
            fecha_produccion=fecha_operacion,
            notas=f"[AUTO_REABASTO_CEDIS:{fecha_operacion.isoformat()}] focus master",
            creado_por=self.user,
        )
        PlanProduccionItem.objects.create(plan=plan, receta=self.receta_sin_empaque, cantidad=Decimal("2"))
        insumo_blocked = Insumo.objects.create(
            nombre="Etiqueta reabasto sin categoria",
            categoria="",
            unidad_base=self.unidad_kg,
            proveedor_principal=self.materia_prima.proveedor_principal,
            activo=True,
            tipo_item=Insumo.TIPO_EMPAQUE,
        )
        CostoInsumo.objects.create(
            insumo=insumo_blocked,
            proveedor=self.materia_prima.proveedor_principal,
            costo_unitario=Decimal("2.20"),
            source_hash="cost-reabasto-master-focus",
        )
        SolicitudCompra.objects.create(
            area=f"PLAN_PRODUCCION:{plan.id}",
            solicitante="admin",
            insumo=insumo_blocked,
            cantidad=Decimal("2.000"),
            fecha_requerida=fecha_operacion,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )

        response = self.client.get(
            reverse("recetas:reabasto_cedis"),
            {
                "fecha": fecha_operacion.isoformat(),
                "master_focus_key": "EMPAQUE",
            },
        )
        self.assertEqual(response.status_code, 200)
        daily_control = response.context["reabasto_daily_control"]
        self.assertEqual(daily_control["selected_master_focus_key"], "EMPAQUE")
        self.assertTrue(daily_control["master_focus_rows"])
        self.assertTrue(all(row["class_label"] == "Empaque" for row in daily_control["master_focus_rows"]))
        self.assertEqual(daily_control["master_summary"]["label"], "Maestro ERP con bloqueos")
        self.assertGreaterEqual(daily_control["master_summary"]["blocked_count"], 1)
        self.assertLess(daily_control["master_summary"]["progress_pct"], 100)
        focus_row = daily_control["master_focus_rows"][0]
        self.assertIn(f"insumo_id={insumo_blocked.id}", focus_row["action_url"])
        self.assertEqual(focus_row["edit_url"], reverse("maestros:insumo_update", args=[insumo_blocked.id]))
        self.assertContains(response, "Vista enfocada")
        self.assertContains(response, "Bloqueo maestro prioritario")
        self.assertContains(response, "Maestro ERP con bloqueos")
        self.assertContains(response, "Editar artículo")
        self.assertContains(response, "kpi-card is-active", html=False)

    def test_reabasto_cedis_blocks_generation_by_critical_master_demand(self):
        fecha_operacion = timezone.localdate()
        plan = PlanProduccion.objects.create(
            nombre=f"CEDIS Reabasto {fecha_operacion.isoformat()} crítico",
            fecha_produccion=fecha_operacion,
            notas=f"[AUTO_REABASTO_CEDIS:{fecha_operacion.isoformat()}] critical master",
            creado_por=self.user,
        )
        insumo_critico = Insumo.objects.create(
            nombre="Caja critica reabasto",
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        receta_critica = Receta.objects.create(
            nombre="Pastel critico reabasto",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Pastel",
            hash_contenido="hash-reabasto-critico-001",
        )
        LineaReceta.objects.create(
            receta=receta_critica,
            posicion=1,
            insumo=insumo_critico,
            insumo_texto=insumo_critico.nombre,
            cantidad=Decimal("1"),
            unidad=self.unidad_kg,
            unidad_texto="kg",
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        PlanProduccionItem.objects.create(plan=plan, receta=receta_critica, cantidad=Decimal("8"))
        VentaHistorica.objects.create(
            receta=receta_critica,
            sucursal=self.sucursal,
            fecha=fecha_operacion - timedelta(days=2),
            cantidad=Decimal("91"),
            tickets=6,
        )

        response = self.client.get(reverse("recetas:reabasto_cedis"), {"fecha": fecha_operacion.isoformat()})
        self.assertEqual(response.status_code, 200)
        daily_control = response.context["reabasto_daily_control"]
        self.assertFalse(daily_control["generation_gate"]["can_generate_plan"])
        self.assertFalse(daily_control["generation_gate"]["can_generate_compras"])
        self.assertTrue(daily_control["master_demand_rows"])
        self.assertContains(response, "Liberación operativa retenida")
        self.assertTrue(any(card["label"] == "Demanda crítica bloqueada" for card in daily_control["control_cards"]))
        plan_row = next(row for row in daily_control["trunk_handoff_rows"] if row["label"] == "Sucursales / Plan")
        self.assertEqual(plan_row["tone"], "danger")
        self.assertEqual(plan_row["status"], "Crítico")
        critical_check = next(item for item in daily_control["generation_gate"]["checks"] if item["label"] == "Maestro crítico del plan cerrado")
        self.assertFalse(critical_check["is_ready"])
        self.assertContains(response, "Demanda crítica bloqueada por maestro")
        self.assertContains(response, "Caja critica reabasto")

    def test_reabasto_cedis_can_focus_enterprise_blocker_group(self):
        fecha_operacion = timezone.localdate()
        response = self.client.get(
            reverse("recetas:reabasto_cedis"),
            {
                "fecha": fecha_operacion.isoformat(),
                "board_focus_key": "sin_inventario",
            },
        )
        self.assertEqual(response.status_code, 200)
        board = response.context["reabasto_enterprise_board"]
        self.assertEqual(board["selected_focus_key"], "sin_inventario")
        self.assertTrue(board["detail_rows"])
        self.assertTrue(all(row["blocker_key"] == "sin_inventario" for row in board["detail_rows"]))
        self.assertContains(response, "Bloqueo operativo enfocado")
        self.assertContains(response, "Vista enfocada")
        self.assertContains(response, "kpi-card is-active", html=False)


class RematchLineasRecetaCommandTests(TestCase):
    def setUp(self):
        self.unidad_kg = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.receta = Receta.objects.create(
            nombre="Receta Rematch Command",
            hash_contenido="hash-rematch-command-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            rendimiento_cantidad=Decimal("1"),
            rendimiento_unidad=self.unidad_kg,
        )
        self.linea_meta = LineaReceta.objects.create(
            receta=self.receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            etapa="",
            insumo=None,
            insumo_texto="Presentación",
            cantidad=None,
            unidad_texto="",
            unidad=None,
            costo_linea_excel=Decimal("2.000000"),
            costo_unitario_snapshot=None,
            match_score=63.6,
            match_method=LineaReceta.MATCH_NONE,
            match_status=LineaReceta.STATUS_REJECTED,
        )
        self.linea_sub = LineaReceta.objects.create(
            receta=self.receta,
            posicion=2,
            tipo_linea=LineaReceta.TIPO_SUBSECCION,
            etapa="Decorado",
            insumo=None,
            insumo_texto="Decorado",
            cantidad=Decimal("0.100000"),
            unidad_texto="kg",
            unidad=self.unidad_kg,
            costo_linea_excel=Decimal("0.500000"),
            costo_unitario_snapshot=None,
            match_score=70.0,
            match_method=LineaReceta.MATCH_FUZZY,
            match_status=LineaReceta.STATUS_NEEDS_REVIEW,
        )

    def test_rematch_dry_run_does_not_modify_lines(self):
        call_command(
            "rematch_lineas_receta",
            "--include-needs-review",
            "--receta",
            "Receta Rematch Command",
            "--limit",
            "50",
        )
        self.linea_meta.refresh_from_db()
        self.linea_sub.refresh_from_db()
        self.assertEqual(self.linea_meta.match_status, LineaReceta.STATUS_REJECTED)
        self.assertEqual(self.linea_meta.match_method, LineaReceta.MATCH_NONE)
        self.assertEqual(self.linea_sub.match_status, LineaReceta.STATUS_NEEDS_REVIEW)

    def test_rematch_apply_auto_approves_meta_and_subsection(self):
        with patch("recetas.utils.costeo_versionado.asegurar_version_costeo"):
            call_command(
                "rematch_lineas_receta",
                "--apply",
                "--include-needs-review",
                "--receta",
                "Receta Rematch Command",
                "--limit",
                "50",
            )
        self.linea_meta.refresh_from_db()
        self.linea_sub.refresh_from_db()
        self.assertEqual(self.linea_meta.match_status, LineaReceta.STATUS_AUTO)
        self.assertEqual(self.linea_meta.match_method, "META_LINEA")
        self.assertEqual(self.linea_sub.match_status, LineaReceta.STATUS_AUTO)
        self.assertEqual(self.linea_sub.match_method, LineaReceta.MATCH_SUBSECTION)


class RecetaCosteoSubseccionTests(TestCase):
    def setUp(self):
        self.unidad_kg = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.insumo = Insumo.objects.create(
            nombre="Harina Costeo Sub",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        self.receta = Receta.objects.create(
            nombre="Pastel QA Sin Duplicar",
            hash_contenido="hash-costeo-sub-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            rendimiento_cantidad=Decimal("1"),
            rendimiento_unidad=self.unidad_kg,
        )
        self.linea_normal = LineaReceta.objects.create(
            receta=self.receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            etapa="Base",
            insumo=self.insumo,
            insumo_texto="Harina Costeo Sub",
            cantidad=Decimal("2.000000"),
            unidad_texto="kg",
            unidad=self.unidad_kg,
            costo_linea_excel=None,
            costo_unitario_snapshot=Decimal("10.000000"),
            match_score=100,
            match_method="MANUAL",
            match_status=LineaReceta.STATUS_AUTO,
        )
        self.linea_sub = LineaReceta.objects.create(
            receta=self.receta,
            posicion=2,
            tipo_linea=LineaReceta.TIPO_SUBSECCION,
            etapa="Decorado",
            insumo=None,
            insumo_texto="Decorado QA",
            cantidad=Decimal("0.100000"),
            unidad_texto="kg",
            unidad=self.unidad_kg,
            costo_linea_excel=Decimal("99.000000"),
            costo_unitario_snapshot=None,
            match_score=100,
            match_method=LineaReceta.MATCH_SUBSECTION,
            match_status=LineaReceta.STATUS_AUTO,
        )

    def test_subsection_does_not_duplicate_recipe_rollup(self):
        self.receta.refresh_from_db()
        self.linea_normal.refresh_from_db()
        self.linea_sub.refresh_from_db()

        self.assertEqual(self.linea_normal.costo_total_estimado, 20.0)
        self.assertIsNone(self.linea_sub.costo_total_estimado)
        self.assertEqual(self.receta.costo_total_estimado_decimal, Decimal("20.000000"))


class RecetaCopyLineasTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_copy_lineas",
            email="admin_copy_lineas@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

        self.unidad_kg = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.insumo = Insumo.objects.create(
            nombre="Harina QA Copy",
            unidad_base=self.unidad_kg,
            activo=True,
        )
        self.receta_destino = Receta.objects.create(
            nombre="Pastel Destino Copy",
            hash_contenido="hash-copy-destino-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        self.receta_origen = Receta.objects.create(
            nombre="Pastel Origen Copy",
            hash_contenido="hash-copy-origen-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        LineaReceta.objects.create(
            receta=self.receta_origen,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            etapa="Base",
            insumo=self.insumo,
            insumo_texto="Harina QA Copy",
            cantidad=Decimal("0.500000"),
            unidad_texto="kg",
            unidad=self.unidad_kg,
            costo_linea_excel=None,
            costo_unitario_snapshot=Decimal("10.000000"),
            match_score=100,
            match_method="MANUAL",
            match_status=LineaReceta.STATUS_AUTO,
            aprobado_por=self.user,
            aprobado_en=timezone.now(),
        )
        LineaReceta.objects.create(
            receta=self.receta_origen,
            posicion=2,
            tipo_linea=LineaReceta.TIPO_SUBSECCION,
            etapa="Decorado",
            insumo=None,
            insumo_texto="Decorado QA Copy",
            cantidad=Decimal("0.100000"),
            unidad_texto="kg",
            unidad=self.unidad_kg,
            costo_linea_excel=Decimal("3.000000"),
            costo_unitario_snapshot=None,
            match_score=100,
            match_method=LineaReceta.MATCH_SUBSECTION,
            match_status=LineaReceta.STATUS_AUTO,
            aprobado_por=self.user,
            aprobado_en=timezone.now(),
        )

    def test_receta_copy_lineas_append_adds_source_lines_at_end(self):
        LineaReceta.objects.create(
            receta=self.receta_destino,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            etapa="Previo",
            insumo=self.insumo,
            insumo_texto="Previo QA Copy",
            cantidad=Decimal("0.200000"),
            unidad_texto="kg",
            unidad=self.unidad_kg,
            costo_unitario_snapshot=Decimal("8.000000"),
            match_score=100,
            match_method="MANUAL",
            match_status=LineaReceta.STATUS_AUTO,
        )

        with patch("recetas.views._sync_derived_insumos_safe"), patch("recetas.views._sync_cost_version_safe"), patch("recetas.views.log_event"):
            response = self.client.post(
                reverse("recetas:receta_copy_lineas", args=[self.receta_destino.id]),
                {"source_receta_id": self.receta_origen.id, "copy_mode": "append"},
                follow=True,
            )

        self.assertEqual(response.status_code, 200)
        lineas = list(self.receta_destino.lineas.order_by("posicion"))
        self.assertEqual(len(lineas), 3)
        self.assertEqual([linea.posicion for linea in lineas], [1, 2, 3])
        self.assertEqual(lineas[1].insumo_texto, "Harina QA Copy")
        self.assertEqual(lineas[2].insumo_texto, "Decorado QA Copy")

    def test_receta_copy_lineas_replace_replaces_existing_lines(self):
        LineaReceta.objects.create(
            receta=self.receta_destino,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            etapa="Vieja",
            insumo=self.insumo,
            insumo_texto="Vieja QA Copy",
            cantidad=Decimal("0.900000"),
            unidad_texto="kg",
            unidad=self.unidad_kg,
            costo_unitario_snapshot=Decimal("9.000000"),
            match_score=100,
            match_method="MANUAL",
            match_status=LineaReceta.STATUS_AUTO,
        )

        with patch("recetas.views._sync_derived_insumos_safe"), patch("recetas.views._sync_cost_version_safe"), patch("recetas.views.log_event"):
            response = self.client.post(
                reverse("recetas:receta_copy_lineas", args=[self.receta_destino.id]),
                {"source_receta_id": self.receta_origen.id, "copy_mode": "replace"},
                follow=True,
            )

        self.assertEqual(response.status_code, 200)
        lineas = list(self.receta_destino.lineas.order_by("posicion"))
        self.assertEqual(len(lineas), 2)
        self.assertEqual([linea.posicion for linea in lineas], [1, 2])
        self.assertEqual(lineas[0].insumo_texto, "Harina QA Copy")
        self.assertEqual(lineas[1].insumo_texto, "Decorado QA Copy")

    def test_receta_copy_lineas_rejects_same_recipe(self):
        with patch("recetas.views._sync_derived_insumos_safe"), patch("recetas.views._sync_cost_version_safe"), patch("recetas.views.log_event"):
            response = self.client.post(
                reverse("recetas:receta_copy_lineas", args=[self.receta_origen.id]),
                {"source_receta_id": self.receta_origen.id, "copy_mode": "append"},
                follow=True,
            )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No puedes copiar ingredientes desde la misma receta.")
        self.assertEqual(self.receta_origen.lineas.count(), 2)


class RecetaCreateWizardTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_receta_create",
            email="admin_receta_create@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)
        self.unidad_kg = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )

    def test_receta_create_view_loads(self):
        response = self.client.get(reverse("recetas:receta_create"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Centro de mando ERP")
        self.assertContains(response, "Workflow ERP del alta")
        self.assertContains(response, "Alta guiada de receta")
        self.assertContains(response, "Insumo interno base")
        self.assertContains(response, "Producto final de venta")
        self.assertContains(response, "Taxonomía enterprise")
        self.assertIn("familia_categoria_catalogo_json", response.context)

    def test_receta_create_creates_producto_final(self):
        response = self.client.post(
            reverse("recetas:receta_create"),
            {
                "recipe_mode": "FINAL",
                "nombre": "Pastel QA Wizard",
                "codigo_point": "PT-QA-001",
                "familia": "Pasteles",
                "categoria": "Chocolate",
                "sheet_name": "Pasteles QA",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        receta = Receta.objects.get(nombre="Pastel QA Wizard")
        self.assertEqual(receta.tipo, Receta.TIPO_PRODUCTO_FINAL)
        self.assertFalse(receta.usa_presentaciones)
        self.assertEqual(receta.familia, "Pasteles")

    def test_receta_create_creates_base_with_presentaciones(self):
        response = self.client.post(
            reverse("recetas:receta_create"),
            {
                "recipe_mode": "BASE_DERIVADOS",
                "nombre": "Pan QA Wizard",
                "categoria": "Vainilla",
                "sheet_name": "Insumos 1",
                "rendimiento_cantidad": "10.4",
                "rendimiento_unidad_id": str(self.unidad_kg.id),
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        receta = Receta.objects.get(nombre="Pan QA Wizard")
        self.assertEqual(receta.tipo, Receta.TIPO_PREPARACION)
        self.assertTrue(receta.usa_presentaciones)
        self.assertEqual(receta.rendimiento_unidad_id, self.unidad_kg.id)

    def test_receta_create_requires_rendimiento_for_base_modes(self):
        response = self.client.post(
            reverse("recetas:receta_create"),
            {
                "recipe_mode": "BASE",
                "nombre": "Batida QA Sin Rendimiento",
                "categoria": "Prueba",
                "sheet_name": "Insumos 1",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Debes capturar el rendimiento total de la batida para costeo enterprise.")
        self.assertContains(response, "Debes seleccionar la unidad del rendimiento para costeo enterprise.")
        self.assertFalse(Receta.objects.filter(nombre="Batida QA Sin Rendimiento").exists())

    def test_receta_create_producto_final_ignores_rendimiento(self):
        response = self.client.post(
            reverse("recetas:receta_create"),
            {
                "recipe_mode": "FINAL",
                "nombre": "Pastel QA Sin Rendimiento",
                "codigo_point": "PT-QA-002",
                "familia": "Pasteles",
                "categoria": "Chocolate",
                "sheet_name": "Pasteles QA",
                "rendimiento_cantidad": "10.5",
                "rendimiento_unidad_id": str(self.unidad_kg.id),
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        receta = Receta.objects.get(nombre="Pastel QA Sin Rendimiento")
        self.assertEqual(receta.tipo, Receta.TIPO_PRODUCTO_FINAL)
        self.assertIsNone(receta.rendimiento_cantidad)
        self.assertIsNone(receta.rendimiento_unidad)

    def test_receta_detail_shows_component_breakdown_for_producto_final(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Breakdown",
            hash_contenido="hash-breakdown-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pasteles",
        )
        mp = Insumo.objects.create(nombre="Fresa QA", tipo_item=Insumo.TIPO_MATERIA_PRIMA, unidad_base=self.unidad_kg, activo=True)
        interno = Insumo.objects.create(nombre="Dream Whip QA", tipo_item=Insumo.TIPO_INTERNO, unidad_base=self.unidad_kg, activo=True)
        empaque = Insumo.objects.create(nombre="Caja QA", tipo_item=Insumo.TIPO_EMPAQUE, unidad_base=self.unidad_kg, activo=True)
        LineaReceta.objects.create(receta=receta, posicion=1, tipo_linea=LineaReceta.TIPO_NORMAL, insumo=interno, insumo_texto=interno.nombre, unidad=self.unidad_kg, unidad_texto="kg", cantidad=Decimal("1.000000"), costo_unitario_snapshot=Decimal("10.000000"), match_status=LineaReceta.STATUS_AUTO, match_method="MANUAL", match_score=100)
        LineaReceta.objects.create(receta=receta, posicion=2, tipo_linea=LineaReceta.TIPO_NORMAL, insumo=mp, insumo_texto=mp.nombre, unidad=self.unidad_kg, unidad_texto="kg", cantidad=Decimal("1.000000"), costo_unitario_snapshot=Decimal("5.000000"), match_status=LineaReceta.STATUS_AUTO, match_method="MANUAL", match_score=100)
        LineaReceta.objects.create(receta=receta, posicion=3, tipo_linea=LineaReceta.TIPO_NORMAL, insumo=empaque, insumo_texto=empaque.nombre, unidad=self.unidad_kg, unidad_texto="kg", cantidad=Decimal("1.000000"), costo_unitario_snapshot=Decimal("2.000000"), match_status=LineaReceta.STATUS_AUTO, match_method="MANUAL", match_score=100)

        response = self.client.get(reverse("recetas:receta_detail", args=[receta.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Composición del costo")
        self.assertContains(response, "Insumos internos")
        self.assertContains(response, "Materia prima puntual")
        self.assertContains(response, "Empaques")
        self.assertContains(response, "+ Agregar interno")
        self.assertContains(response, "+ Agregar MP")
        self.assertContains(response, "+ Agregar empaque")

    def test_receta_detail_edit_links_preserve_component_context(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Links",
            hash_contenido="hash-breakdown-001c",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pasteles",
        )
        interno = Insumo.objects.create(
            nombre="Relleno QA Link",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad_kg,
            activo=True,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=interno,
            insumo_texto=interno.nombre,
            unidad=self.unidad_kg,
            unidad_texto="kg",
            cantidad=Decimal("1.000000"),
            costo_unitario_snapshot=Decimal("12.500000"),
            match_status=LineaReceta.STATUS_AUTO,
            match_method="MANUAL",
            match_score=100,
        )

        response = self.client.get(reverse("recetas:receta_detail", args=[receta.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "component_context=internos")

    def test_receta_detail_producto_final_marks_rendimiento_as_not_applicable(self):
        receta = Receta.objects.create(
            nombre="Pastel QA Sin Rendimiento Visible",
            hash_contenido="hash-breakdown-001b",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pasteles",
        )
        response = self.client.get(reverse("recetas:receta_detail", args=[receta.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Taxonomía enterprise")
        self.assertContains(response, "Rendimiento no aplica")
        self.assertContains(response, "Costo por rendimiento no aplica")


class RecetasListEnterpriseChainTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_recipe_chain",
            email="admin_recipe_chain@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)
        self.unidad = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.unidad_pza = UnidadMedida.objects.create(
            codigo="pza",
            nombre="Pieza",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        self.proveedor = Proveedor.objects.create(nombre="Proveedor Chain")

    def test_recetas_list_shows_supply_chain_summary_for_base_recipe(self):
        base = Receta.objects.create(
            nombre="Base Lista Cadena",
            hash_contenido="hash-recetas-list-chain-001",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("10"),
            rendimiento_unidad=self.unidad,
        )
        derivado = Insumo.objects.create(
            nombre="Base Lista Cadena - Chico",
            codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:CHICO",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )
        final = Receta.objects.create(
            nombre="Pastel Lista Cadena",
            hash_contenido="hash-recetas-list-chain-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=derivado,
            insumo_texto=derivado.nombre,
            cantidad=Decimal("2"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("4"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "insumos"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Derivados:")
        self.assertContains(response, "Finales ligados:")
        self.assertContains(response, "Pastel Lista Cadena")

    def test_recetas_list_shows_chain_checkpoints_for_base_recipe(self):
        base = Receta.objects.create(
            nombre="Base Checkpoint Cadena",
            hash_contenido="hash-recetas-list-chain-001b",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("8"),
            rendimiento_unidad=self.unidad,
        )
        presentacion_chico = RecetaPresentacion.objects.create(
            receta=base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.500000"),
            activo=True,
        )
        derivado = Insumo.objects.create(
            nombre="Base Checkpoint Cadena - Chico",
            codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:CHICO",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )
        final = Receta.objects.create(
            nombre="Pastel Checkpoint Cadena",
            hash_contenido="hash-recetas-list-chain-001c",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=derivado,
            insumo_texto=derivado.nombre,
            cantidad=Decimal("1"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("4"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "subinsumos", "q": base.nombre})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Base")
        self.assertContains(response, "Derivados")
        self.assertContains(response, "Uso final")
        self.assertContains(response, "Derivados")

    def test_recetas_list_shows_enterprise_stage_for_base_recipe(self):
        base = Receta.objects.create(
            nombre="Base Etapa ERP",
            hash_contenido="hash-recetas-list-stage-001",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("8"),
            rendimiento_unidad=self.unidad,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "subinsumos", "q": base.nombre})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Etapa ERP:")
        self.assertContains(response, "Derivados en configuración")

    def test_recetas_list_shows_enterprise_stage_playbook(self):
        base = Receta.objects.create(
            nombre="Base Playbook ERP",
            hash_contenido="hash-recetas-list-stage-playbook-001",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("8"),
            rendimiento_unidad=self.unidad,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "subinsumos", "q": base.nombre})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Playbook de etapa")
        self.assertContains(response, "Resumen de etapa")
        self.assertContains(response, "Presentaciones y derivados")
        self.assertContains(response, "Sincronizar derivados")
        self.assertContains(response, "Cierre:")

    def test_recetas_list_can_filter_by_enterprise_stage(self):
        base = Receta.objects.create(
            nombre="Base Filtro Etapa ERP",
            hash_contenido="hash-recetas-list-stage-filter-001",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("8"),
            rendimiento_unidad=self.unidad,
        )

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "subinsumos", "enterprise_stage": "derivados_setup"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Etapa ERP: Derivados en configuración")
        self.assertContains(response, base.nombre)

    def test_recetas_list_can_filter_bases_without_downstream_consumption(self):
        base = Receta.objects.create(
            nombre="Base Sin Consumo Final",
            hash_contenido="hash-recetas-list-chain-003",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("6"),
            rendimiento_unidad=self.unidad,
        )
        presentacion_chico = RecetaPresentacion.objects.create(
            receta=base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.500000"),
            activo=True,
        )
        Insumo.objects.create(
            nombre="Base Sin Consumo Final - Chico",
            codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:CHICO",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "subinsumos", "governance_issue": "sin_consumo_final"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sin consumo final")
        self.assertContains(response, "Base Sin Consumo Final")

    def test_recetas_list_can_filter_products_without_base_origin(self):
        interno = Insumo.objects.create(
            nombre="Interno Lista Sin Origen",
            codigo="INT-LISTA-SIN-ORIGEN",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )
        final = Receta.objects.create(
            nombre="Pastel Lista Sin Origen",
            hash_contenido="hash-recetas-list-chain-004",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "productos", "governance_issue": "sin_base_origen"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sin base origen")
        self.assertContains(response, "Pastel Lista Sin Origen")

    def test_recetas_list_can_filter_products_without_packaging(self):
        interno = Insumo.objects.create(
            nombre="Interno Lista Sin Empaque",
            codigo="INT-LISTA-SIN-EMPAQUE",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Rellenos",
        )
        final = Receta.objects.create(
            nombre="Pastel Lista Sin Empaque",
            hash_contenido="hash-recetas-list-chain-004-b",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "productos", "governance_issue": "sin_empaque"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sin empaque")
        self.assertContains(response, "Pastel Lista Sin Empaque")
        self.assertContains(response, "Agregar empaque")

    def test_recetas_list_shows_primary_action_for_missing_packaging(self):
        interno = Insumo.objects.create(
            nombre="Interno Acción Sin Empaque",
            codigo="INT-ACCION-SIN-EMPAQUE",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Rellenos",
        )
        final = Receta.objects.create(
            nombre="Pastel Acción Sin Empaque",
            hash_contenido="hash-recetas-list-chain-004-d",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos", "q": "Acción Sin Empaque"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Agregar empaque")

    def test_recetas_list_marks_missing_packaging_as_operational_warning(self):
        interno = Insumo.objects.create(
            nombre="Interno Health Sin Empaque",
            codigo="INT-HEALTH-SIN-EMPAQUE",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Rellenos",
        )
        final = Receta.objects.create(
            nombre="Pastel Health Sin Empaque",
            hash_contenido="hash-recetas-list-chain-004-h",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos", "q": "Health Sin Empaque"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sin empaque")
        self.assertContains(response, "Producto final todavía sin empaque ligado en su BOM.")

    def test_recetas_list_does_not_flag_galleta_without_fixed_packaging(self):
        materia_prima = Insumo.objects.create(
            nombre="Chocolate Galleta Flexible",
            codigo="MP-GALLETA-FLEXIBLE",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unidad,
            activo=True,
            categoria="Chocolate",
        )
        final = Receta.objects.create(
            nombre="Galleta Flexible Operativa",
            hash_contenido="hash-recetas-list-chain-004-flex-galleta",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Galletas",
            categoria="Galletas",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=materia_prima,
            insumo_texto=materia_prima.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "productos", "governance_issue": "sin_empaque"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Galleta Flexible Operativa")

    def test_recetas_list_marks_galleta_without_fixed_packaging_as_ready(self):
        materia_prima = Insumo.objects.create(
            nombre="Chocolate Galleta Health",
            codigo="MP-GALLETA-HEALTH",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unidad,
            activo=True,
            categoria="Chocolate",
        )
        final = Receta.objects.create(
            nombre="Galleta Health Operativa",
            hash_contenido="hash-recetas-list-chain-004-flex-health",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Galletas",
            categoria="Galletas",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=materia_prima,
            insumo_texto=materia_prima.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos", "q": "Health Operativa"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Empaque flexible")
        self.assertNotContains(response, "Producto final todavía sin empaque ligado en su BOM.")

    def test_recetas_list_does_not_flag_empanada_without_fixed_packaging(self):
        interno = Insumo.objects.create(
            nombre="Relleno Empanada Flexible",
            codigo="INT-EMPANADA-FLEXIBLE",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Rellenos",
        )
        final = Receta.objects.create(
            nombre="Empanada Flexible Operativa",
            hash_contenido="hash-recetas-list-chain-004-flex-empanada",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Empanadas",
            categoria="Empanadas",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "productos", "governance_issue": "sin_empaque"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Empanada Flexible Operativa")

    def test_recetas_list_marks_empanada_without_fixed_packaging_as_ready(self):
        interno = Insumo.objects.create(
            nombre="Relleno Empanada Health",
            codigo="INT-EMPANADA-HEALTH",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Rellenos",
        )
        final = Receta.objects.create(
            nombre="Empanada Health Operativa",
            hash_contenido="hash-recetas-list-chain-004-flex-empanada-health",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Empanadas",
            categoria="Empanadas",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos", "q": "Empanada Health Operativa"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Empaque flexible")
        self.assertNotContains(response, "Producto final todavía sin empaque ligado en su BOM.")

    def test_recetas_list_marks_bolitas_kg_without_fixed_packaging_as_ready(self):
        interno = Insumo.objects.create(
            nombre="Masa Bolitas Health",
            codigo="INT-BOLITAS-HEALTH",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Masas",
        )
        final = Receta.objects.create(
            nombre="Bolitas de Nuez KG",
            codigo_point="05021",
            hash_contenido="hash-recetas-list-chain-004-flex-bolitas-health",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Galletas",
            categoria="Galletas",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos", "q": "Bolitas de Nuez KG"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Empaque flexible")
        self.assertNotContains(response, "Producto final todavía sin empaque ligado en su BOM.")

    def test_recetas_list_shows_internal_components_checkpoint_card(self):
        empaque = Insumo.objects.create(
            nombre="Caja Checkpoint Sin Internos",
            codigo="EMP-CHECKPOINT-SIN-INTERNOS",
            tipo_item=Insumo.TIPO_EMPAQUE,
            unidad_base=self.unidad,
            activo=True,
            categoria="Empaques",
        )
        final = Receta.objects.create(
            nombre="Pastel Checkpoint Sin Internos",
            hash_contenido="hash-recetas-list-chain-004-internals",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=empaque,
            insumo_texto=empaque.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="pza",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos", "q": "Sin Internos"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sin internos")
        self.assertContains(response, "sin insumos internos suficientes en la estructura")

    def test_recetas_list_shows_chain_checkpoints_for_final_recipe(self):
        interno = Insumo.objects.create(
            nombre="Interno Checkpoint Final",
            codigo="INT-CHECKPOINT-FINAL",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Rellenos",
        )
        empaque = Insumo.objects.create(
            nombre="Caja Checkpoint Final",
            codigo="EMP-CHECKPOINT-FINAL",
            tipo_item=Insumo.TIPO_EMPAQUE,
            unidad_base=self.unidad,
            activo=True,
            categoria="Empaques",
        )
        final = Receta.objects.create(
            nombre="Pastel Checkpoint Final",
            hash_contenido="hash-recetas-list-chain-004-z",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=2,
            insumo=empaque,
            insumo_texto=empaque.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="pza",
            costo_unitario_snapshot=Decimal("2"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos", "q": final.nombre})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Internos")
        self.assertContains(response, "Trazabilidad")
        self.assertContains(response, "Empaque")
        self.assertContains(response, "1 ligado(s)")

    def test_recetas_list_can_filter_by_chain_status_pending(self):
        base = Receta.objects.create(
            nombre="Base Cadena Pendiente Filter",
            hash_contenido="hash-recetas-list-chain-007",
            tipo=Receta.TIPO_PREPARACION,
            familia="Bases",
            categoria="Chocolate",
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("4.000000"),
            rendimiento_unidad=self.unidad,
        )
        presentacion_chico = RecetaPresentacion.objects.create(
            receta=base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.500000"),
            activo=True,
        )

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "subinsumos", "chain_status": "pendientes", "q": base.nombre},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cadena ERP")
        self.assertContains(response, "Derivados por actualizar")
        self.assertContains(response, "Base Cadena Pendiente Filter")

    def test_recetas_list_can_filter_by_chain_checkpoint_final_usage(self):
        base = Receta.objects.create(
            nombre="Base Checkpoint Sin Consumo",
            hash_contenido="hash-recetas-list-chain-007b",
            tipo=Receta.TIPO_PREPARACION,
            familia="Bases",
            categoria="Chocolate",
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("4.000000"),
            rendimiento_unidad=self.unidad,
        )
        presentacion_chico = RecetaPresentacion.objects.create(
            receta=base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.500000"),
            activo=True,
        )
        Insumo.objects.create(
            nombre="Base Checkpoint Sin Consumo - Chico",
            codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:CHICO",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "subinsumos", "chain_checkpoint": "final_usage"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sin consumo final")
        self.assertContains(response, "Base Checkpoint Sin Consumo")

    def test_recetas_list_can_filter_by_chain_checkpoint_packaging_ready(self):
        interno = Insumo.objects.create(
            nombre="Interno Checkpoint Empaque Filtro",
            codigo="INT-CHECKPOINT-EMPAQUE-FILTRO",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Rellenos",
        )
        final = Receta.objects.create(
            nombre="Pastel Checkpoint Empaque Filtro",
            hash_contenido="hash-recetas-list-chain-007c",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(
            reverse("recetas:recetas_list"),
            {"vista": "productos", "chain_checkpoint": "packaging_ready"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sin empaque")
        self.assertContains(response, "Pastel Checkpoint Empaque Filtro")

    def test_recetas_list_shows_chain_focus_for_products_without_packaging(self):
        base = Receta.objects.create(
            nombre="Base Focus Producto Empaque",
            hash_contenido="hash-recetas-list-chain-focus-prod-base",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("6.000000"),
            rendimiento_unidad=self.unidad,
        )
        interno = Insumo.objects.create(
            nombre="Interno Focus Sin Empaque",
            codigo=f"DERIVADO:RECETA:{base.id}:PREPARACION",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Rellenos",
        )
        final = Receta.objects.create(
            nombre="Pastel Focus Sin Empaque",
            hash_contenido="hash-recetas-list-chain-focus-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["chain_focus"]["checkpoint"], "packaging_ready")
        self.assertEqual(response.context["chain_focus"]["action_label"], "Agregar empaques")
        self.assertContains(response, "Qué falta y qué hacer")

        self.assertContains(response, "Cadena ERP prioritaria")
        self.assertContains(response, "Empaque faltante")
        self.assertContains(response, "Agregar empaques")

    def test_recetas_list_shows_chain_focus_for_bases_without_sync(self):
        base = Receta.objects.create(
            nombre="Base Focus Sync Pendiente",
            hash_contenido="hash-recetas-list-chain-focus-002",
            tipo=Receta.TIPO_PREPARACION,
            familia="Bases",
            categoria="Chocolate",
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("6.000000"),
            rendimiento_unidad=self.unidad,
        )
        presentacion_chico = RecetaPresentacion.objects.create(
            receta=base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.500000"),
            activo=True,
        )
        presentacion_mediano = RecetaPresentacion.objects.create(
            receta=base,
            nombre="Mediano",
            peso_por_unidad_kg=Decimal("0.800000"),
            activo=True,
        )
        derivado = Insumo.objects.get(codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:{presentacion_chico.id}")
        Insumo.objects.filter(codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:{presentacion_mediano.id}").update(
            activo=False
        )
        final = Receta.objects.create(
            nombre="Pastel Focus Sync Pendiente",
            hash_contenido="hash-recetas-list-chain-focus-final",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=derivado,
            insumo_texto=derivado.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "subinsumos"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["chain_focus"]["checkpoint"], "derived_sync")
        self.assertEqual(response.context["chain_focus"]["action_label"], "Sincronizar derivados")
        self.assertContains(response, "Sincronizar derivados")

    def test_receta_detail_shows_action_for_missing_packaging(self):
        interno = Insumo.objects.create(
            nombre="Interno Detalle Sin Empaque",
            codigo="INT-DETALLE-SIN-EMPAQUE",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Rellenos",
        )
        final = Receta.objects.create(
            nombre="Pastel Detalle Sin Empaque",
            hash_contenido="hash-recetas-list-chain-004-c",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:receta_detail", args=[final.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sin empaque ligado")
        self.assertContains(response, "Agregar empaque")

    def test_receta_detail_shows_master_blockers_panel(self):
        interno = Insumo.objects.create(
            nombre="Interno Detalle Maestro Incompleto",
            codigo="INT-DETALLE-MAESTRO",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="",
        )
        final = Receta.objects.create(
            nombre="Pastel Detalle Maestro",
            hash_contenido="hash-recetas-list-chain-004-maestro",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:receta_detail", args=[final.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Brechas del maestro")
        self.assertContains(response, "Interno Detalle Maestro Incompleto")
        self.assertContains(response, "Falta: categoría")
        self.assertContains(response, "Abrir artículo")
        self.assertContains(response, "missing_field=categoria")
        self.assertContains(response, f"linked_recipe_id={final.id}")
        self.assertContains(response, "impact_scope=finales")

    def test_receta_detail_shows_action_for_base_without_presentaciones(self):
        base = Receta.objects.create(
            nombre="Base Detalle Sin Presentaciones",
            hash_contenido="hash-recetas-list-chain-004-e",
            tipo=Receta.TIPO_PREPARACION,
            familia="Bases",
            categoria="Chocolate",
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("5.000000"),
            rendimiento_unidad=self.unidad,
        )

        response = self.client.get(reverse("recetas:receta_detail", args=[base.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sin presentaciones activas")
        self.assertContains(response, "Agregar presentación")

    def test_receta_detail_shows_action_for_base_without_final_consumption(self):
        base = Receta.objects.create(
            nombre="Base Detalle Sync Pendiente",
            hash_contenido="hash-recetas-list-chain-004-sync",
            tipo=Receta.TIPO_PREPARACION,
            familia="Bases",
            categoria="Chocolate",
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("5.000000"),
            rendimiento_unidad=self.unidad,
        )
        RecetaPresentacion.objects.create(
            receta=base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.500000"),
            activo=True,
        )

        response = self.client.get(reverse("recetas:receta_detail", args=[base.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sin consumo final")
        self.assertContains(response, "Crear producto final")

    def test_receta_detail_shows_chain_focus_summary(self):
        base = Receta.objects.create(
            nombre="Base Focus",
            hash_contenido="hash-base-focus-001",
            tipo=Receta.TIPO_PREPARACION,
            familia="Bases",
            categoria="Pan",
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("4.000000"),
            rendimiento_unidad=self.unidad,
        )

        response = self.client.get(reverse("recetas:receta_detail", args=[base.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Bloqueo dominante de cadena")
        self.assertContains(response, "Crear presentaciones")

    def test_receta_detail_shows_operational_handoff_table(self):
        receta = Receta.objects.create(
            nombre="Pastel Handoff ERP",
            hash_contenido="hash-pastel-handoff-erp-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        interno = Insumo.objects.create(
            nombre="Ganache Handoff",
            tipo_item=Insumo.TIPO_INTERNO,
            categoria="Relleno",
            unidad_base=self.unidad,
            activo=True,
        )
        empaque = Insumo.objects.create(
            nombre="Caja Handoff",
            tipo_item=Insumo.TIPO_EMPAQUE,
            categoria="Empaque",
            unidad_base=self.unidad_pza,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=interno,
            proveedor=self.proveedor,
            costo_unitario=Decimal("10.00"),
            source_hash="handoff-interno-cost-001",
        )
        CostoInsumo.objects.create(
            insumo=empaque,
            proveedor=self.proveedor,
            costo_unitario=Decimal("5.00"),
            source_hash="handoff-empaque-cost-001",
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=2,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=empaque,
            insumo_texto=empaque.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad_pza,
            unidad_texto="pza",
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:receta_detail", args=[receta.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Salida operativa del documento")
        self.assertContains(response, "Cierre troncal ERP consolidado")
        self.assertContains(response, "Dependencias del flujo")
        self.assertContains(response, "Costeo")
        self.assertContains(response, "Recetas / BOM")
        self.assertContains(response, "Compras")
        self.assertContains(response, "Compras documentales")
        self.assertContains(response, "Inventario / Reabasto")
        self.assertContains(response, "Criterio de salida")
        self.assertIn("trunk_handoff_rows", response.context)
        self.assertEqual(len(response.context["trunk_handoff_rows"]), 3)

    def test_receta_detail_shows_enterprise_stage(self):
        base = Receta.objects.create(
            nombre="Base Etapa Detalle",
            hash_contenido="hash-base-stage-detail-001",
            tipo=Receta.TIPO_PREPARACION,
            familia="Bases",
            categoria="Pan",
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("4.000000"),
            rendimiento_unidad=self.unidad,
        )

        response = self.client.get(reverse("recetas:receta_detail", args=[base.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Etapa de cierre")
        self.assertContains(response, "Derivados en configuración")
        self.assertContains(response, "Resumen de etapa")
        self.assertContains(response, "Presentaciones y derivados")
        self.assertContains(response, "Sincronizar derivados")

    def test_recetas_list_shows_primary_action_for_base_without_presentaciones(self):
        base = Receta.objects.create(
            nombre="Base Acción Sin Presentaciones",
            hash_contenido="hash-recetas-list-chain-004-f",
            tipo=Receta.TIPO_PREPARACION,
            familia="Bases",
            categoria="Chocolate",
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("5.000000"),
            rendimiento_unidad=self.unidad,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "subinsumos", "q": base.nombre})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Agregar presentación")

    def test_recetas_list_shows_primary_action_for_base_without_final_consumption(self):
        base = Receta.objects.create(
            nombre="Base Acción Sin Consumo Final",
            hash_contenido="hash-recetas-list-chain-004-g",
            tipo=Receta.TIPO_PREPARACION,
            familia="Bases",
            categoria="Chocolate",
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("6.000000"),
            rendimiento_unidad=self.unidad,
        )
        RecetaPresentacion.objects.create(
            receta=base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.500000"),
            activo=True,
        )
        Insumo.objects.create(
            nombre="Base Acción Sin Consumo Final - Chico",
            codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:CHICO",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Bases",
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "subinsumos", "q": base.nombre})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Crear producto final")
        self.assertContains(response, f"source_base={base.id}")

    def test_recetas_list_shows_chain_focus_summary_for_producto_final_without_empaque(self):
        receta = Receta.objects.create(
            nombre="Pastel Focus",
            hash_contenido="hash-producto-focus-empaque-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
            sheet_name="Pasteles",
        )
        interno = Insumo.objects.create(
            nombre="Ganache Focus",
            tipo_item=Insumo.TIPO_INTERNO,
            categoria="Relleno",
            unidad_base=self.unidad,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=interno,
            proveedor=self.proveedor,
            costo_unitario=Decimal("10.00"),
            source_hash="focus-interno-cost-001",
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos", "q": receta.nombre})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Qué falta y qué hacer")
        self.assertContains(response, "Agregar empaque")

    def test_recetas_list_shows_downstream_handoff_table(self):
        receta = Receta.objects.create(
            nombre="Pastel Downstream ERP",
            hash_contenido="hash-pastel-downstream-erp-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        interno = Insumo.objects.create(
            nombre="Ganache Downstream",
            tipo_item=Insumo.TIPO_INTERNO,
            categoria="Relleno",
            unidad_base=self.unidad,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=interno,
            proveedor=self.proveedor,
            costo_unitario=Decimal("9.50"),
            source_hash="downstream-interno-cost-001",
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=interno,
            insumo_texto=interno.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos", "q": receta.nombre})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Qué falta y qué hacer")
        self.assertContains(response, "Abrir ficha")
        self.assertIn("trunk_handoff_rows", response.context)

    def test_recetas_list_shows_chain_action_links_for_base_pending_sync(self):
        base = Receta.objects.create(
            nombre="Base Acción Sync Derivados",
            hash_contenido="hash-recetas-list-chain-004-h",
            tipo=Receta.TIPO_PREPARACION,
            familia="Bases",
            categoria="Chocolate",
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("6.000000"),
            rendimiento_unidad=self.unidad,
        )
        RecetaPresentacion.objects.create(
            receta=base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.500000"),
            activo=True,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "subinsumos", "q": base.nombre})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Presentaciones")
        self.assertContains(response, "Derivados por actualizar")
        self.assertContains(response, reverse("recetas:presentacion_create", args=[base.id]))

    def test_recetas_list_shows_upstream_summary_for_product_final(self):
        base = Receta.objects.create(
            nombre="Base Lista Origen",
            hash_contenido="hash-recetas-list-chain-005",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("7"),
            rendimiento_unidad=self.unidad,
        )
        derivado = Insumo.objects.create(
            nombre="Base Lista Origen - Chico",
            codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:CHICO",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
        )
        final = Receta.objects.create(
            nombre="Pastel Lista Origen",
            hash_contenido="hash-recetas-list-chain-006",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=derivado,
            insumo_texto=derivado.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Bases origen:")
        self.assertContains(response, "Internos sin origen:")
        self.assertContains(response, "Base Lista Origen")

    def test_recetas_list_shows_apply_suggested_action_for_direct_base_with_match(self):
        base = Receta.objects.create(
            nombre="Base Lista Directa Acción",
            hash_contenido="hash-recetas-list-directa-accion-001",
            tipo=Receta.TIPO_PREPARACION,
            familia="Bases",
            categoria="Chocolate",
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("10.000000"),
            rendimiento_unidad=self.unidad,
        )
        base_directa = Insumo.objects.create(
            nombre="Base Lista Directa Acción",
            codigo=f"DERIVADO:RECETA:{base.id}:PREPARACION",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Bases",
        )
        empaque = Insumo.objects.create(
            nombre="Caja Lista Directa Acción",
            codigo="EMP-LISTA-DIRECTA-ACCION",
            tipo_item=Insumo.TIPO_EMPAQUE,
            unidad_base=self.unidad,
            activo=True,
            categoria="Empaques",
        )
        presentacion = RecetaPresentacion.objects.create(
            receta=base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.500000"),
            activo=True,
        )
        derivado = Insumo.objects.create(
            nombre="Base Lista Directa Acción - Chico",
            codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:{presentacion.id}",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Bases",
        )
        final = Receta.objects.create(
            nombre="Pastel Lista Directa Acción",
            hash_contenido="hash-recetas-list-directa-accion-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=base_directa,
            insumo_texto=base_directa.nombre,
            cantidad=Decimal("0.500000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=2,
            insumo=empaque,
            insumo_texto=empaque.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="pza",
            costo_unitario_snapshot=Decimal("1"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos", "q": final.nombre})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Aplicar derivados sugeridos")
        self.assertContains(response, "Sugerencias listas:")
        self.assertContains(response, derivado.nombre)
        self.assertContains(response, reverse("recetas:receta_apply_direct_base_replacements", args=[final.id]))

    def test_recetas_list_keeps_direct_base_review_action_when_no_suggestion_exists(self):
        base = Receta.objects.create(
            nombre="Base Lista Directa Sin Sugerencia",
            hash_contenido="hash-recetas-list-directa-sin-sugerencia-001",
            tipo=Receta.TIPO_PREPARACION,
            familia="Bases",
            categoria="Chocolate",
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("10.000000"),
            rendimiento_unidad=self.unidad,
        )
        base_directa = Insumo.objects.create(
            nombre="Base Lista Directa Sin Sugerencia",
            codigo=f"DERIVADO:RECETA:{base.id}:PREPARACION",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Bases",
        )
        RecetaPresentacion.objects.create(
            receta=base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.500000"),
            activo=True,
        )
        final = Receta.objects.create(
            nombre="Pastel Lista Directa Sin Sugerencia",
            hash_contenido="hash-recetas-list-directa-sin-sugerencia-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=base_directa,
            insumo_texto=base_directa.nombre,
            cantidad=Decimal("0.500000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos", "q": final.nombre})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ajustar componente")
        self.assertNotContains(response, "Aplicar derivados sugeridos")

    def test_recetas_list_ignores_implausible_direct_base_replacement(self):
        base = Receta.objects.create(
            nombre="Base Lista Directa Implausible",
            hash_contenido="hash-recetas-list-directa-implausible-001",
            tipo=Receta.TIPO_PREPARACION,
            familia="Bases",
            categoria="Chocolate",
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("10.000000"),
            rendimiento_unidad=self.unidad,
        )
        base_directa = Insumo.objects.create(
            nombre="Base Lista Directa Implausible",
            codigo=f"DERIVADO:RECETA:{base.id}:PREPARACION",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Bases",
        )
        presentacion = RecetaPresentacion.objects.create(
            receta=base,
            nombre="Individual",
            peso_por_unidad_kg=Decimal("0.090000"),
            activo=True,
        )
        Insumo.objects.create(
            nombre="Base Lista Directa Implausible - Individual",
            codigo=f"DERIVADO:RECETA:{base.id}:PRESENTACION:{presentacion.id}",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
            activo=True,
            categoria="Bases",
        )
        empaque = Insumo.objects.create(
            nombre="Caja Lista Directa Implausible",
            codigo="EMP-LISTA-DIRECTA-IMPLAUSIBLE",
            tipo_item=Insumo.TIPO_EMPAQUE,
            unidad_base=self.unidad,
            activo=True,
            categoria="Empaques",
        )
        final = Receta.objects.create(
            nombre="Pastel Lista Directa Implausible",
            hash_contenido="hash-recetas-list-directa-implausible-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Chocolate",
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=1,
            insumo=base_directa,
            insumo_texto=base_directa.nombre,
            cantidad=Decimal("8.000000"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("5"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        LineaReceta.objects.create(
            receta=final,
            posicion=2,
            insumo=empaque,
            insumo_texto=empaque.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="pza",
            costo_unitario_snapshot=Decimal("1"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.get(reverse("recetas:recetas_list"), {"vista": "productos", "q": final.nombre})

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Usa base sin presentación")
        self.assertNotContains(response, "Aplicar derivados sugeridos")


class RecetaDeleteViewTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_recetas_delete",
            email="admin_recetas_delete@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)
        self.unidad = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.sucursal = Sucursal.objects.create(codigo="MTZ", nombre="Matriz")

    def test_delete_receta_succeeds_without_operational_footprint(self):
        receta = Receta.objects.create(
            nombre="Receta Borrable",
            hash_contenido="hash-receta-borrable",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("5"),
            rendimiento_unidad=self.unidad,
        )

        response = self.client.post(reverse("recetas:receta_delete", args=[receta.id]))

        self.assertRedirects(response, reverse("recetas:recetas_list"))
        self.assertFalse(Receta.objects.filter(pk=receta.id).exists())

    def test_delete_receta_is_blocked_when_has_sales_history(self):
        receta = Receta.objects.create(
            nombre="Receta Con Venta",
            hash_contenido="hash-receta-con-venta",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        VentaHistorica.objects.create(
            receta=receta,
            sucursal=self.sucursal,
            fecha=date(2026, 3, 20),
            cantidad=Decimal("4"),
            tickets=2,
            monto_total=Decimal("250"),
        )

        response = self.client.post(reverse("recetas:receta_delete", args=[receta.id]), follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertTrue(Receta.objects.filter(pk=receta.id).exists())
        self.assertContains(response, "huella operativa")


class MrpRapidoEnterpriseTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="mrpadmin",
            email="mrpadmin@example.com",
            password="testpass123",
        )
        self.client.force_login(self.user)
        self.unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo")
        self.insumo = Insumo.objects.create(
            nombre="Chocolate prueba MRP",
            activo=True,
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unidad,
        )
        CostoInsumo.objects.create(
            insumo=self.insumo,
            costo_unitario=Decimal("25"),
            source_hash="mrp-rapido-enterprise-base-cost-001",
        )
        self.receta = Receta.objects.create(
            nombre="MRP Rápido Chocolate",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Pastel",
        )
        LineaReceta.objects.create(
            receta=self.receta,
            posicion=1,
            insumo=self.insumo,
            insumo_texto=self.insumo.nombre,
            cantidad=Decimal("2"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("25"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        self.sucursal_mrp = Sucursal.objects.create(codigo="SUC-MRP-CRIT", nombre="Sucursal MRP Critica", activa=True)
        VentaHistorica.objects.create(
            receta=self.receta,
            sucursal=self.sucursal_mrp,
            fecha=timezone.localdate() - timedelta(days=2),
            cantidad=Decimal("88"),
            tickets=7,
        )
        for offset, qty in ((6, "21"), (10, "18"), (14, "16"), (20, "13")):
            VentaHistorica.objects.create(
                receta=self.receta,
                sucursal=self.sucursal_mrp,
                fecha=timezone.localdate() - timedelta(days=offset),
                cantidad=Decimal(qty),
                tickets=3,
            )

    def test_mrp_rapido_renders_enterprise_health(self):
        response = self.client.post(
            reverse("recetas:mrp_form"),
            {"receta_id": self.receta.id, "multiplicador": "1"},
        )
        self.assertEqual(response.status_code, 200)
        resultado = response.context["resultado"]
        self.assertIn("erp_command_center", response.context)
        self.assertIn("workflow_rows", response.context)
        self.assertIn("critical_path_rows", response.context)
        self.assertEqual(resultado["health_label"], "Atención operativa")
        self.assertEqual(resultado["alertas_capacidad"], 1)
        self.assertEqual(resultado["master_incompletos"], 1)
        self.assertTrue(any(card["label"] == "Stock insuficiente" for card in resultado["quality_cards"]))
        self.assertTrue(
            any(card["label"] in {"Maestro incompleto", "Demanda crítica bloqueada por maestro"} for card in resultado["quality_cards"])
        )
        self.assertTrue(resultado["master_blocker_class_cards"])
        self.assertEqual(resultado["master_blocker_class_cards"][0]["class_label"], "Materia prima")
        self.assertEqual(resultado["master_focus"]["class_label"], "Materia prima")
        self.assertEqual(resultado["master_focus"]["missing_field"], "proveedor")
        self.assertTrue(resultado["master_focus_rows"])
        self.assertTrue(resultado["downstream_handoff_rows"])
        focus_row = resultado["master_focus_rows"][0]
        self.assertIn(f"insumo_id={self.insumo.id}", focus_row["action_url"])
        self.assertEqual(focus_row["edit_url"], reverse("maestros:insumo_update", args=[self.insumo.id]))
        self.assertTrue(
            any(row["missing"] == "proveedor principal, código comercial" for row in resultado["master_blocker_detail_rows"])
        )
        self.assertContains(response, "Resumen del cálculo")
        self.assertContains(response, "Flujo del cálculo")
        self.assertContains(response, "Entrega del cálculo a downstream")
        self.assertContains(response, "Prioridades de atención")
        self.assertContains(response, "Control comercial:")
        self.assertContains(response, "Bloqueos del cálculo")
        self.assertContains(response, "Brechas del maestro")
        self.assertContains(response, "Bloqueo maestro prioritario")
        self.assertContains(response, "Cadena ERP afectada")
        self.assertContains(response, "Editar artículo")
        self.assertContains(response, "Compras")
        self.assertContains(response, "Inventario")
        self.assertContains(response, "Stock insuficiente")
        self.assertContains(response, "Liberación operativa retenida")
        self.assertContains(response, "Faltante: proveedor principal, código comercial")
        self.assertIn("critical_master_rows", resultado)
        self.assertTrue(resultado["critical_master_rows"])
        self.assertEqual(response.context["erp_command_center"]["tone"], "danger")
        self.assertEqual(response.context["erp_command_center"]["status"], "Crítico")
        master_row = next(row for row in response.context["workflow_rows"] if row["title"] == "Maestro del artículo")
        self.assertEqual(master_row["tone"], "danger")
        self.assertTrue(
            any(row["label"] == "Maestro crítico por demanda" for row in resultado["upstream_dependency_rows"])
        )
        self.assertTrue(
            any(
                card["label"] in {"Demanda crítica bloqueada por maestro", "Maestro bloqueando MRP"}
                for card in resultado["chain_cards"]
            )
        )
        self.assertIn("commercial_gate", resultado)

    def test_mrp_rapido_marks_base_directa_and_non_canonical(self):
        receta_base = Receta.objects.create(
            nombre="Base MRP Chocolate",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
            rendimiento_cantidad=Decimal("10"),
            rendimiento_unidad=self.unidad,
            hash_contenido="hash-mrp-base-directa-001",
        )
        Insumo.objects.create(
            nombre="BASE MRP CHOCOLATE",
            codigo=f"DERIVADO:RECETA:{receta_base.id}:PREPARACION:CANON",
            codigo_point="MRP-BASE-001",
            activo=True,
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
        )
        base_directa = Insumo.objects.create(
            nombre="Base MRP Chocolate",
            codigo=f"DERIVADO:RECETA:{receta_base.id}:PREPARACION",
            activo=True,
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
        )
        presentacion = RecetaPresentacion.objects.create(
            receta=receta_base,
            nombre="Chico",
            peso_por_unidad_kg=Decimal("0.50"),
            activo=True,
        )
        derivado = Insumo.objects.create(
            nombre="Base MRP Chocolate - Chico",
            codigo=f"DERIVADO:RECETA:{receta_base.id}:PRESENTACION:{presentacion.id}",
            activo=True,
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unidad,
        )
        CostoInsumo.objects.create(
            insumo=derivado,
            costo_unitario=Decimal("12"),
            source_hash="mrp-rapido-enterprise-derivado-cost-001",
        )

        receta = Receta.objects.create(
            nombre="Producto MRP directo",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Pastel",
            hash_contenido="hash-mrp-final-directa-001",
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=base_directa,
            insumo_texto=base_directa.nombre,
            cantidad=Decimal("1"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("10"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.post(reverse("recetas:mrp_form"), {"receta_id": receta.id, "multiplicador": "1"})
        self.assertEqual(response.status_code, 200)
        resultado = response.context["resultado"]
        self.assertEqual(resultado["lineas_base_directa"], 1)
        self.assertEqual(resultado["lineas_base_directa_sugeridas"], 1)
        self.assertEqual(resultado["lineas_no_canonicas"], 1)
        self.assertEqual(resultado["health_label"], "Con bloqueos operativos")
        self.assertTrue(any(card["label"] == "Usa base sin presentación" for card in resultado["quality_cards"]))
        self.assertTrue(any(card["label"] == "Fuera de estándar" for card in resultado["quality_cards"]))
        self.assertTrue(any(card["label"] == "Bases sin presentación detectadas" for card in resultado["chain_cards"]))
        self.assertTrue(
            any(
                card["label"] == "Usa base sin presentación"
                and card["action_label"] == "Aplicar derivados sugeridos"
                and card.get("action_method") == "post"
                for card in resultado["quality_cards"]
            )
        )
        self.assertContains(response, "Artículo propuesto")
        self.assertContains(response, "Usa una base completa")
        self.assertContains(response, "Cadena ERP afectada")
        self.assertContains(response, "Base origen")
        self.assertContains(response, "Base MRP Chocolate - Chico")
        self.assertContains(response, "Es la presentación activa más cercana")
        self.assertContains(response, "Aplicar derivados sugeridos")

    def test_mrp_rapido_can_focus_quality_card(self):
        response = self.client.post(
            reverse("recetas:mrp_form"),
            {
                "receta_id": self.receta.id,
                "multiplicador": "1",
                "focus_kind": "quality",
                "focus_key": "maestro_incompleto",
            },
        )
        self.assertEqual(response.status_code, 200)
        resultado = response.context["resultado"]
        self.assertEqual(resultado["selected_focus_kind"], "quality")
        self.assertEqual(resultado["selected_focus_key"], "maestro_incompleto")
        self.assertEqual(len(resultado["items"]), 1)
        self.assertEqual(resultado["items"][0]["nombre"], self.insumo.nombre)
        self.assertContains(response, "Vista enfocada")
        self.assertContains(response, "kpi-card is-active", html=False)

    def test_mrp_rapido_can_focus_chain_and_master_cards(self):
        empaque = Insumo.objects.create(
            nombre="Caja MRP",
            activo=True,
            tipo_item=Insumo.TIPO_EMPAQUE,
            unidad_base=self.unidad,
            categoria="Empaque",
        )
        CostoInsumo.objects.create(
            insumo=empaque,
            costo_unitario=Decimal("4"),
            source_hash="mrp-rapido-empaque-cost-001",
        )
        LineaReceta.objects.create(
            receta=self.receta,
            posicion=2,
            insumo=empaque,
            insumo_texto=empaque.nombre,
            cantidad=Decimal("1"),
            unidad=self.unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("4"),
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )

        response = self.client.post(
            reverse("recetas:mrp_form"),
            {
                "receta_id": self.receta.id,
                "multiplicador": "1",
                "focus_kind": "master",
                "focus_key": Insumo.TIPO_MATERIA_PRIMA,
            },
        )
        self.assertEqual(response.status_code, 200)
        resultado = response.context["resultado"]
        self.assertEqual(resultado["selected_focus_kind"], "master")
        self.assertEqual(resultado["selected_focus_key"], Insumo.TIPO_MATERIA_PRIMA)
        self.assertEqual(len(resultado["items"]), 1)
        self.assertTrue(all(item["origen"] == "Materia prima" for item in resultado["items"]))

        response = self.client.post(
            reverse("recetas:mrp_form"),
            {
                "receta_id": self.receta.id,
                "multiplicador": "1",
                "focus_kind": "chain",
                "focus_key": "maestro_bloqueando",
            },
        )
        self.assertEqual(response.status_code, 200)
        resultado = response.context["resultado"]
        self.assertEqual(resultado["selected_focus_kind"], "chain")
        self.assertEqual(resultado["selected_focus_key"], "maestro_bloqueando")
        self.assertEqual(len(resultado["items"]), 2)
        self.assertTrue(all(item["master_missing"] for item in resultado["items"]))
        self.assertEqual(resultado["chain_detail_rows"][0]["type"], "Maestro incompleto")

    def test_mrp_rapido_renders_recipe_demand_signal(self):
        sucursal = Sucursal.objects.create(codigo="001", nombre="Centro", activa=True)
        today = timezone.localdate()
        VentaHistorica.objects.create(
            receta=self.receta,
            sucursal=sucursal,
            fecha=today - timedelta(days=2),
            cantidad=Decimal("12"),
            tickets=4,
        )
        VentaHistorica.objects.create(
            receta=self.receta,
            sucursal=sucursal,
            fecha=today - timedelta(days=9),
            cantidad=Decimal("10"),
            tickets=3,
        )
        periodo = today.strftime("%Y-%m")
        month_start = date(today.year, today.month, 1)
        month_end = date(today.year, today.month, monthrange(today.year, today.month)[1])
        PronosticoVenta.objects.create(
            receta=self.receta,
            periodo=periodo,
            cantidad=Decimal("20"),
            fuente="TEST",
        )
        SolicitudVenta.objects.create(
            receta=self.receta,
            sucursal=None,
            alcance=SolicitudVenta.ALCANCE_MES,
            periodo=periodo,
            fecha_inicio=month_start,
            fecha_fin=month_end,
            cantidad=Decimal("20"),
            fuente="TEST",
        )
        session = self.client.session
        session["pronostico_estadistico_preview"] = {
            "alcance": "mes",
            "periodo": periodo,
            "target_start": month_start.isoformat(),
            "target_end": month_end.isoformat(),
            "sucursal_id": None,
            "sucursal_nombre": "Todas",
            "rows": [
                {
                    "receta_id": self.receta.id,
                    "receta": self.receta.nombre,
                    "forecast_qty": 20.0,
                    "forecast_low": 18.0,
                    "forecast_high": 22.0,
                    "desviacion": 2.0,
                    "muestras": 12,
                    "pronostico_actual": 20.0,
                    "delta": 0.0,
                    "recomendacion": "MANTENER",
                    "observaciones": 12,
                    "confianza": 78.0,
                }
            ],
            "totals": {
                "recetas_count": 1,
                "forecast_total": 20.0,
                "forecast_low_total": 18.0,
                "forecast_high_total": 22.0,
                "pronostico_total": 20.0,
                "delta_total": 0.0,
            },
        }
        session.save()

        response = self.client.post(
            reverse("recetas:mrp_form"),
            {"receta_id": self.receta.id, "multiplicador": "1"},
        )
        self.assertEqual(response.status_code, 200)
        resultado = response.context["resultado"]
        self.assertIn("demand_signal", resultado)
        self.assertIn("upstream_dependency_rows", resultado)
        self.assertEqual(resultado["demand_signal"]["historico_days"], 6)
        self.assertIn("years_observed", resultado["demand_signal"])
        self.assertIn("comparable_years", resultado["demand_signal"])
        self.assertEqual(resultado["demand_signal"]["forecast_status"], "Forecast confiable")
        self.assertEqual(resultado["demand_signal"]["alignment_status"], "Solicitud alineada")
        self.assertContains(response, "Señal comercial de la receta")
        self.assertContains(response, "Años observados")
        self.assertContains(response, "Dependencias previas del cálculo")
        self.assertContains(response, "Demanda comercial")
        self.assertContains(response, "Alineación forecast/solicitud")
        self.assertContains(response, "Forecast vigente")


class CosteoCiruelaGuardrailTests(TestCase):
    def setUp(self):
        self.kg = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.g = UnidadMedida.objects.create(
            codigo="g",
            nombre="Gramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1"),
        )
        self.lt = UnidadMedida.objects.create(
            codigo="lt",
            nombre="Litro",
            tipo=UnidadMedida.TIPO_VOLUMEN,
            factor_to_base=Decimal("1000"),
        )
        self.ml = UnidadMedida.objects.create(
            codigo="ml",
            nombre="Mililitro",
            tipo=UnidadMedida.TIPO_VOLUMEN,
            factor_to_base=Decimal("1"),
        )
        self.pza = UnidadMedida.objects.create(
            codigo="pza",
            nombre="Pieza",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )

    def test_guardrail_forces_protected_ciruela_yield(self):
        from recetas.utils.rendimientos_protegidos import enforce_protected_preparation_yield

        receta = Receta.objects.create(
            nombre="Ciruela Cocida",
            codigo_point="01CC07",
            hash_contenido="ciruela-cocida-guardrail",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("1"),
            rendimiento_unidad=self.kg,
        )

        qty, unit, protected = enforce_protected_preparation_yield(receta, Decimal("1"), self.kg)

        self.assertTrue(protected)
        self.assertEqual(qty, Decimal("5.172000"))
        self.assertEqual(unit, self.kg)

    def test_restaurar_costeo_ciruela_restores_yields_and_versions(self):
        insumo_ciruela = Insumo.objects.create(nombre="Ciruela", unidad_base=self.kg, activo=True)
        CostoInsumo.objects.create(
            insumo=insumo_ciruela,
            costo_unitario=Decimal("119.47"),
            source_hash="ciruela-cost-guardrail",
        )
        insumo_agua = Insumo.objects.create(nombre="AGUA", unidad_base=self.lt, activo=True)
        CostoInsumo.objects.create(
            insumo=insumo_agua,
            costo_unitario=Decimal("0.895"),
            source_hash="agua-cost-guardrail",
        )
        insumo_azucar = Insumo.objects.create(nombre="AZUCAR ESTANDAR", unidad_base=self.kg, activo=True)
        CostoInsumo.objects.create(
            insumo=insumo_azucar,
            costo_unitario=Decimal("18.80"),
            source_hash="azucar-cost-guardrail",
        )

        ciruela_cocida = Receta.objects.create(
            nombre="Ciruela Cocida",
            codigo_point="01CC07",
            hash_contenido="ciruela-cocida-command",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("1"),
            rendimiento_unidad=self.kg,
        )
        jugo = Receta.objects.create(
            nombre="Jugo de Ciruela",
            codigo_point="03JUC64",
            hash_contenido="jugo-ciruela-command",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("1"),
            rendimiento_unidad=self.lt,
        )
        mermelada = Receta.objects.create(
            nombre="Mermelada de Ciruela",
            codigo_point="02MC08",
            hash_contenido="mermelada-ciruela-command",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("1"),
            rendimiento_unidad=self.kg,
        )
        final_recipes = [
            Receta.objects.create(
                nombre="Pastel de Ciruela Chico",
                codigo_point="0113",
                hash_contenido="pastel-ciruela-chico-command",
                tipo=Receta.TIPO_PRODUCTO_FINAL,
            ),
            Receta.objects.create(
                nombre="Pastel de Ciruela Mediano",
                codigo_point="0112",
                hash_contenido="pastel-ciruela-mediano-command",
                tipo=Receta.TIPO_PRODUCTO_FINAL,
            ),
            Receta.objects.create(
                nombre="Pastel de Ciruela Grande",
                codigo_point="0111",
                hash_contenido="pastel-ciruela-grande-command",
                tipo=Receta.TIPO_PRODUCTO_FINAL,
            ),
            Receta.objects.create(
                nombre="Pastel de Ciruela R",
                codigo_point="0114",
                hash_contenido="pastel-ciruela-r-command",
                tipo=Receta.TIPO_PRODUCTO_FINAL,
            ),
        ]

        LineaReceta.objects.create(
            receta=ciruela_cocida,
            posicion=1,
            insumo=insumo_ciruela,
            insumo_texto="CIRUELA",
            cantidad=Decimal("2260"),
            unidad=self.g,
            unidad_texto="g",
            match_status=LineaReceta.STATUS_AUTO,
        )
        LineaReceta.objects.create(
            receta=ciruela_cocida,
            posicion=2,
            insumo=insumo_agua,
            insumo_texto="AGUA",
            cantidad=Decimal("5204"),
            unidad=self.ml,
            unidad_texto="ml",
            match_status=LineaReceta.STATUS_AUTO,
        )
        LineaReceta.objects.create(
            receta=ciruela_cocida,
            posicion=3,
            insumo=insumo_azucar,
            insumo_texto="AZUCAR ESTANDAR",
            cantidad=Decimal("4000"),
            unidad=self.g,
            unidad_texto="g",
            match_status=LineaReceta.STATUS_AUTO,
        )

        call_command(
            "restaurar_costeo_ciruela",
            "--anchor-date=2026-06-02",
            "--skip-historical",
            stdout=StringIO(),
            verbosity=0,
        )

        ciruela_cocida.refresh_from_db()
        jugo.refresh_from_db()
        mermelada.refresh_from_db()
        self.assertEqual(ciruela_cocida.rendimiento_cantidad, Decimal("5.172000"))
        self.assertEqual(jugo.rendimiento_cantidad, Decimal("1.078000"))
        self.assertEqual(mermelada.rendimiento_cantidad, Decimal("4.094000"))
        self.assertTrue(RecetaCostoVersion.objects.filter(receta=ciruela_cocida, fuente="RESTAURA_RENDIMIENTO_CIRUELA").exists())
        for receta in final_recipes:
            self.assertTrue(RecetaCostoVersion.objects.filter(receta=receta, fuente="RESTAURA_RENDIMIENTO_CIRUELA").exists())
