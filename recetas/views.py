import hashlib
import csv
from io import BytesIO
from math import sqrt
from datetime import date, datetime, timedelta
from calendar import monthrange
from collections import defaultdict
from decimal import Decimal, ROUND_CEILING
from typing import Dict, Any, List
from urllib.parse import urlencode
from uuid import uuid4

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.exceptions import PermissionDenied
from django.db import OperationalError, ProgrammingError, transaction
from django.db.models.deletion import ProtectedError
from django.db.models import Count, Q, OuterRef, Subquery, Case, When, Value, IntegerField, Sum, DecimalField, Max
from django.core.paginator import Paginator
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.urls import reverse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from rapidfuzz import fuzz

from compras.models import OrdenCompra, RecepcionCompra, SolicitudCompra
from control.models import MermaPOS
from core.access import can_manage_compras, can_view_recetas, is_branch_capture_only
from core.audit import log_event
from core.models import Sucursal, sucursales_operativas
from inventario.models import ExistenciaInsumo, MovimientoInventario
from maestros.models import CostoInsumo, Insumo, UnidadMedida
from pos_bridge.config import load_point_bridge_settings
from pos_bridge.models import (
    PointDailyBranchIndicator,
    PointDailySale,
    PointInventorySnapshot,
    PointSyncJob,
    PointProductionLine,
    PointTransferLine,
    PointWasteLine,
)
from pos_bridge.services.product_recipe_sync_service import PointProductRecipeSyncService
from maestros.utils.canonical_catalog import (
    canonical_insumo,
    canonical_insumo_by_id,
    canonicalized_insumo_selector,
    latest_costo_canonico,
)
from .models import (
    Receta,
    RecetaAgrupacionAddon,
    RecetaCodigoPointAlias,
    RecetaCostoSemanal,
    normalizar_codigo_point,
    LineaReceta,
    RecetaPresentacion,
    RecetaCostoVersion,
    CostoDriver,
    PlanProduccion,
    PlanProduccionItem,
    PronosticoVenta,
    SolicitudVenta,
    VentaHistorica,
    PoliticaStockSucursalProducto,
    InventarioCedisProducto,
    MovimientoProductoCedis,
    SolicitudReabastoCedis,
    SolicitudReabastoCedisLinea,
)
from .utils.costeo_versionado import asegurar_version_costeo, calcular_costeo_receta, comparativo_versiones
from .utils.costeo_semanal import snapshot_weekly_costs, week_bounds
from .utils.costeo_snapshot import resolve_insumo_unit_cost, resolve_line_snapshot_cost
from .utils.derived_product_presentations import (
    build_upstream_snapshot as build_derived_product_upstream_snapshot,
    get_active_derived_relation,
)
from .utils.derived_insumos import sync_presentacion_insumo, sync_receta_derivados
from .utils.matching import match_insumo
from .utils.normalizacion import normalizar_nombre
from .catalogs import familia_categoria_catalogo_json, familias_producto_catalogo
from reportes.executive_panels import _partial_month_amount_quantity

OFFICIAL_POINT_SOURCE = "/Report/PrintReportes?idreporte=3"
RECENT_POINT_SOURCE = "/Report/VentasCategorias"


def _presentacion_sort_key(nombre: str) -> tuple[int, str]:
    norm = normalizar_nombre(nombre or "")
    order = {
        "mini": 10,
        "chico": 20,
        "mediano": 30,
        "grande": 40,
        "bollos": 50,
        "bollo": 50,
        "bollito": 50,
        "individual": 60,
        "media plancha": 70,
        "1 2 plancha": 70,
        "rosca": 80,
    }
    for token, rank in order.items():
        if f" {token}" in f" {norm}" or norm.endswith(token):
            return rank, norm
    return 999, norm


def _insumo_canonical_priority(insumo: Insumo) -> int:
    score = 0
    if (insumo.codigo_point or "").strip():
        score += 100
    latest_cost = getattr(insumo, "latest_costo_unitario", None)
    if latest_cost is not None and Decimal(str(latest_cost or 0)) > 0:
        score += 60
    if insumo.proveedor_principal_id:
        score += 30
    if insumo.unidad_base_id:
        score += 20
    if (insumo.codigo or "").strip():
        score += 10
    return score


def _canonicalize_insumo_match(insumo: Insumo | None) -> Insumo | None:
    return canonical_insumo(insumo)


def _insumo_erp_readiness(insumo: Insumo) -> dict[str, object]:
    missing: list[str] = []
    codigo_norm = (insumo.codigo or "").strip().upper()
    is_recipe_derived = codigo_norm.startswith("DERIVADO:RECETA:")
    effective_tipo = insumo.tipo_item
    if is_recipe_derived:
        effective_tipo = Insumo.TIPO_INTERNO
    if not insumo.activo:
        missing.append("inactivo")
    if not insumo.unidad_base_id:
        missing.append("unidad base")
    if effective_tipo == Insumo.TIPO_MATERIA_PRIMA and not insumo.proveedor_principal_id:
        missing.append("proveedor principal")
    if (
        effective_tipo == Insumo.TIPO_INTERNO
        and not is_recipe_derived
        and not (insumo.categoria or "").strip()
    ):
        missing.append("categoría")
    if effective_tipo == Insumo.TIPO_EMPAQUE and not (insumo.categoria or "").strip():
        missing.append("categoría")
    return {
        "ready": not missing,
        "missing": missing,
        "label": "Listo para operar" if not missing else "Incompleto",
    }


def _insumo_display_name(insumo: Insumo | None) -> str:
    if not insumo:
        return ""
    return (insumo.nombre_point or insumo.nombre or "").strip()


def _insumo_article_class(insumo: Insumo) -> dict[str, str]:
    if insumo.tipo_item == Insumo.TIPO_EMPAQUE:
        return {
            "key": Insumo.TIPO_EMPAQUE,
            "label": "Empaque",
        }
    if insumo.tipo_item == Insumo.TIPO_INTERNO or (insumo.codigo or "").startswith("DERIVADO:RECETA:"):
        return {
            "key": Insumo.TIPO_INTERNO,
            "label": "Insumo interno",
        }
    return {
        "key": Insumo.TIPO_MATERIA_PRIMA,
        "label": "Materia prima",
    }


def _filtered_recipe_sync_queryset(params) -> Any:
    vista = (params.get("vista") or "").strip().lower()
    q = (params.get("q") or "").strip()
    tipo = (params.get("tipo") or "").strip().upper()
    modo_operativo = (params.get("modo_operativo") or "").strip().upper()
    familia = (params.get("familia") or "").strip()
    categoria = (params.get("categoria") or "").strip()

    qs = Receta.objects.all().order_by("nombre", "id")
    if vista == "productos":
        qs = qs.filter(tipo=Receta.TIPO_PRODUCTO_FINAL)
    elif vista == "insumos":
        qs = qs.filter(tipo=Receta.TIPO_PREPARACION)
    elif vista == "subinsumos":
        qs = qs.filter(tipo=Receta.TIPO_PREPARACION, usa_presentaciones=True)

    if tipo in {Receta.TIPO_PREPARACION, Receta.TIPO_PRODUCTO_FINAL}:
        qs = qs.filter(tipo=tipo)
    if modo_operativo == "BASE":
        qs = qs.filter(tipo=Receta.TIPO_PREPARACION, usa_presentaciones=False)
    elif modo_operativo == "BASE_DERIVADOS":
        qs = qs.filter(tipo=Receta.TIPO_PREPARACION, usa_presentaciones=True)
    elif modo_operativo == "FINAL":
        qs = qs.filter(tipo=Receta.TIPO_PRODUCTO_FINAL)
    if familia:
        qs = qs.filter(familia=familia)
    if categoria:
        qs = qs.filter(categoria=categoria)
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(codigo_point__icontains=q))
    return qs


def _run_point_recipe_sync_action(
    request: HttpRequest,
    *,
    action_label: str,
    product_codes: list[str] | None = None,
    branch_hint: str = "MATRIZ",
    include_without_recipe: bool = False,
) -> tuple[PointSyncJob, dict[str, Any]]:
    clean_codes = [str(code).strip().upper() for code in (product_codes or []) if str(code).strip()]
    job = PointSyncJob.objects.create(
        job_type=PointSyncJob.JOB_TYPE_RECIPES,
        status=PointSyncJob.STATUS_RUNNING,
        parameters={
            "action": action_label,
            "branch_hint": branch_hint,
            "product_codes": clean_codes,
            "include_without_recipe": include_without_recipe,
        },
        triggered_by=request.user,
    )
    service = PointProductRecipeSyncService()
    try:
        result = service.sync(
            branch_hint=branch_hint,
            product_codes=clean_codes or None,
            include_without_recipe=include_without_recipe,
            sync_job=job,
        )
        job.status = PointSyncJob.STATUS_SUCCESS
        job.finished_at = timezone.now()
        job.result_summary = result.summary
        job.artifacts = {"raw_export_path": result.raw_export_path}
        job.save(update_fields=["status", "finished_at", "result_summary", "artifacts", "updated_at"])
        log_event(
            request.user,
            "SYNC_POINT_RECIPES",
            "pos_bridge.PointSyncJob",
            job.id,
            {
                "action": action_label,
                "product_codes": clean_codes,
                "summary": result.summary,
                "raw_export_path": result.raw_export_path,
            },
        )
        return job, result.summary
    except Exception as exc:
        job.status = PointSyncJob.STATUS_FAILED
        job.finished_at = timezone.now()
        job.error_message = str(exc)
        job.save(update_fields=["status", "finished_at", "error_message", "updated_at"])
        log_event(
            request.user,
            "SYNC_POINT_RECIPES_FAILED",
            "pos_bridge.PointSyncJob",
            job.id,
            {
                "action": action_label,
                "product_codes": clean_codes,
                "error": str(exc),
            },
        )
        raise


def _snapshot_costeo_after_sync(*, receta_ids: list[int] | None = None) -> dict[str, Any]:
    summary = snapshot_weekly_costs(receta_ids=receta_ids)
    return {
        "week_start": summary.week_start.isoformat(),
        "week_end": summary.week_end.isoformat(),
        "recipes_created": summary.recipes_created,
        "recipes_updated": summary.recipes_updated,
        "addons_created": summary.addons_created,
        "addons_updated": summary.addons_updated,
        "total_items": summary.total_items,
    }


def _missing_field_to_filter_key(missing_label: str) -> str | None:
    mapping = {
        "unidad base": "unidad",
        "proveedor principal": "proveedor",
        "categoría": "categoria",
        "código comercial": "codigo_point",
        "codigo comercial": "codigo_point",
        "código externo": "codigo_point",
        "codigo externo": "codigo_point",
    }
    return mapping.get(missing_label)


def _enterprise_blocker_action_meta_for_recipes(
    item_name: str,
    class_key: str,
    missing_field: str | None,
    insumo_id: int | None = None,
    usage_scope: str = "recipes",
) -> dict[str, str]:
    missing_value = (missing_field or "").strip()
    normalized = normalizar_nombre(missing_value)
    query = {
        "tipo_item": class_key,
        "enterprise_status": "incompletos",
        "usage_scope": usage_scope,
    }
    filter_key = _missing_field_to_filter_key(missing_value) if missing_value else None
    if filter_key:
        query["missing_field"] = filter_key
    if insumo_id:
        query["insumo_id"] = int(insumo_id)
    elif item_name:
        query["q"] = item_name
    list_url = reverse("maestros:insumo_list") + f"?{urlencode(query)}"
    edit_url = reverse("maestros:insumo_update", args=[insumo_id]) if insumo_id else ""

    if "unidad" in normalized:
        return {
            "label": "Definir unidad base",
            "detail": "Asigna la unidad base para que costeo, recetas y abastecimiento trabajen con la misma unidad operativa.",
            "url": list_url,
            "edit_url": edit_url,
        }
    if "proveedor" in normalized:
        return {
            "label": "Asignar proveedor principal",
            "detail": "Define el proveedor principal para habilitar compras, lead time y reposición del artículo.",
            "url": list_url,
            "edit_url": edit_url,
        }
    if "categoria" in normalized:
        return {
            "label": "Asignar categoría",
            "detail": "Completa la categoría para clasificar el artículo correctamente en recetas, compras y reportes.",
            "url": list_url,
            "edit_url": edit_url,
        }
    if "codigo point" in normalized or "point" in normalized or "codigo comercial" in normalized or "código comercial" in normalized:
        return {
            "label": "Registrar código comercial",
            "detail": "Registra el código comercial para conciliar ventas, surtido y catálogo operativo.",
            "url": list_url,
            "edit_url": edit_url,
        }
    if "costo" in normalized:
        return {
            "label": "Cargar costo vigente",
            "detail": "Registra un costo vigente para que costeo, MRP y compras calculen valores reales.",
            "url": list_url,
            "edit_url": edit_url,
        }
    if "inactivo" in normalized:
        return {
            "label": "Reactivar artículo",
            "detail": "Reactiva el artículo o sustitúyelo por su versión maestra antes de usarlo en recetas o abastecimiento.",
            "url": list_url,
            "edit_url": edit_url,
        }
    if "catalogo" in normalized or "canonico" in normalized:
        return {
            "label": "Revisar catálogo",
            "detail": "Consolida variantes y usa el artículo estándar para evitar duplicidad en costeo, compras y MRP.",
            "url": list_url,
            "edit_url": edit_url,
        }
    return {
        "label": "Abrir maestro",
        "detail": "Completa el maestro del artículo para liberar costeo, MRP, recetas y compras.",
        "url": list_url,
        "edit_url": edit_url,
    }


def _enterprise_blocker_label_detail_for_missing_recipes(missing_field: str | None) -> tuple[str, str]:
    meta = _enterprise_blocker_action_meta_for_recipes("", "", missing_field, usage_scope="recipes")
    return meta["label"], meta["detail"]


def _attach_linea_suggested_match(linea: LineaReceta) -> None:
    linea.suggested_insumo = None
    linea.suggested_score = 0.0
    linea.suggested_method = "NO_MATCH"
    linea.suggested_can_approve = False
    linea.suggested_is_canonical = False
    if linea.tipo_linea == LineaReceta.TIPO_SUBSECCION or linea.insumo_id:
        return
    if linea.match_status not in {LineaReceta.STATUS_NEEDS_REVIEW, LineaReceta.STATUS_REJECTED}:
        return
    if not (linea.insumo_texto or "").strip():
        return

    raw_matched, raw_score, raw_method = match_insumo(linea.insumo_texto)
    suggested_insumo = _canonicalize_insumo_match(raw_matched)
    linea.suggested_insumo = suggested_insumo
    linea.suggested_score = float(raw_score or 0.0)
    linea.suggested_method = raw_method
    linea.suggested_can_approve = bool(suggested_insumo and float(raw_score or 0.0) >= 75.0)
    linea.suggested_is_canonical = bool(
        suggested_insumo and raw_matched and suggested_insumo.id != raw_matched.id
    )


def _attach_linea_canonical_target(linea: LineaReceta) -> None:
    linea.canonical_target = None
    linea.canonical_needs_repoint = False
    if not linea.insumo_id:
        return
    canonical = _canonicalize_insumo_match(linea.insumo)
    if canonical and canonical.id != linea.insumo_id:
        linea.canonical_target = canonical
        linea.canonical_needs_repoint = True


def _validate_selected_insumo_enterprise_ready(linea: LineaReceta) -> str | None:
    if not linea.insumo_id:
        return None
    profile = _insumo_erp_readiness(linea.insumo)
    blocking_missing = [item for item in profile["missing"] if item != "proveedor principal"]
    if not blocking_missing:
        return None
    if blocking_missing == ["inactivo"]:
        return f"El artículo seleccionado ({linea.insumo.nombre}) está inactivo. Usa un artículo vigente."
    missing_human = ", ".join(blocking_missing)
    return (
        f"El artículo seleccionado ({linea.insumo.nombre}) no está listo para operar en ERP. "
        f"Corrige en Maestros: {missing_human}."
    )


def _recipe_search_score(query_norm: str, receta: Receta) -> float:
    """Score de coincidencia aproximada para búsqueda en listado de recetas."""
    if not query_norm:
        return 0.0

    fields = [
        receta.nombre or "",
        receta.codigo_point or "",
        receta.familia or "",
        receta.categoria or "",
        receta.sheet_name or "",
    ]

    best = 0.0
    for field in fields:
        field_norm = normalizar_nombre(field)
        if not field_norm:
            continue
        if query_norm == field_norm:
            return 100.0
        if query_norm in field_norm:
            best = max(best, 96.0)
            continue
        score = max(
            float(fuzz.WRatio(query_norm, field_norm)),
            float(fuzz.partial_ratio(query_norm, field_norm)),
            float(fuzz.token_set_ratio(query_norm, field_norm)),
        )
        best = max(best, score)
    return best


def _recipe_derived_sync_state(receta: Receta) -> dict[str, int | bool]:
    cached = getattr(receta, "_derived_sync_state_cache", None)
    if cached is not None:
        return cached

    prep_code = f"DERIVADO:RECETA:{receta.id}:PREPARACION"
    presentacion_prefix = f"DERIVADO:RECETA:{receta.id}:PRESENTACION:"
    presentaciones_activas = getattr(receta, "presentaciones_activas_count", None)
    if presentaciones_activas is None:
        presentaciones_activas = receta.presentaciones.filter(activo=True).count()

    state = {
        "prep_ready": Insumo.objects.filter(codigo=prep_code, activo=True).exists(),
        "active_presentaciones": int(presentaciones_activas or 0),
        "derived_presentaciones": Insumo.objects.filter(
            codigo__startswith=presentacion_prefix,
            activo=True,
        ).count(),
    }
    setattr(receta, "_derived_sync_state_cache", state)
    return state


def _recipe_supply_chain_snapshot(receta: Receta) -> dict[str, object] | None:
    if receta.tipo != Receta.TIPO_PREPARACION:
        return None

    cached = getattr(receta, "_supply_chain_snapshot_cache", None)
    if cached is not None:
        return cached

    prep_code = f"DERIVADO:RECETA:{receta.id}:PREPARACION"
    presentacion_prefix = f"DERIVADO:RECETA:{receta.id}:PRESENTACION:"

    prep_insumo = (
        Insumo.objects.filter(codigo=prep_code, activo=True)
        .only("id", "nombre", "codigo")
        .first()
    )
    presentacion_qs = Insumo.objects.filter(
        codigo__startswith=presentacion_prefix,
        activo=True,
    ).only("id", "nombre", "codigo")
    presentacion_items = [
        {"id": item.id, "nombre": item.nombre}
        for item in presentacion_qs.order_by("nombre")
    ]
    presentacion_ids = list(presentacion_qs.values_list("id", flat=True))
    component_ids = []
    if prep_insumo:
        component_ids.append(prep_insumo.id)
    component_ids.extend(presentacion_ids)

    downstream_line_qs = LineaReceta.objects.filter(
        receta__tipo=Receta.TIPO_PRODUCTO_FINAL,
        insumo_id__in=component_ids,
    ).select_related("receta", "insumo")
    downstream_recipes = list(
        downstream_line_qs.order_by("receta__nombre")
        .values("receta_id", "receta__nombre")
        .distinct()
    )
    snapshot = {
        "prep_insumo": prep_insumo,
        "prep_insumo_item": (
            {"id": prep_insumo.id, "nombre": prep_insumo.nombre}
            if prep_insumo
            else None
        ),
        "presentacion_count": len(presentacion_ids),
        "presentacion_items": presentacion_items[:8],
        "component_count": len(component_ids),
        "downstream_line_count": downstream_line_qs.count(),
        "downstream_recipe_count": len(downstream_recipes),
        "downstream_recipes": [
            {"id": item["receta_id"], "nombre": item["receta__nombre"]}
            for item in downstream_recipes[:6]
        ],
        "has_downstream_usage": bool(downstream_recipes),
    }
    setattr(receta, "_supply_chain_snapshot_cache", snapshot)
    return snapshot


def _recipe_id_from_derived_code(codigo: str) -> int | None:
    raw = (codigo or "").strip()
    if not raw.startswith("DERIVADO:RECETA:"):
        return None
    parts = raw.split(":")
    if len(parts) < 3:
        return None
    try:
        return int(parts[2])
    except (TypeError, ValueError):
        return None


def _derived_code_kind(codigo: str) -> str | None:
    raw = (codigo or "").strip()
    if not raw.startswith("DERIVADO:RECETA:"):
        return None
    parts = raw.split(":")
    if len(parts) < 4:
        return None
    if parts[3] == "PREPARACION":
        return "PREPARACION"
    if parts[3] == "PRESENTACION":
        return "PRESENTACION"
    return None


def _product_upstream_snapshot(lineas: list[LineaReceta], *, receta: Receta | None = None) -> dict[str, object]:
    source_map: dict[int, dict[str, object]] = {}
    internal_count = 0
    internal_without_source_count = 0
    empaque_count = 0
    mp_count = 0
    for linea in lineas:
        if not linea.insumo_id:
            continue
        if linea.insumo.tipo_item == Insumo.TIPO_EMPAQUE:
            empaque_count += 1
        elif linea.insumo.tipo_item == Insumo.TIPO_MATERIA_PRIMA:
            mp_count += 1
        else:
            internal_count += 1

        source_recipe_id = _recipe_id_from_derived_code(linea.insumo.codigo or "")
        if not source_recipe_id:
            if linea.insumo.tipo_item != Insumo.TIPO_EMPAQUE and linea.insumo.tipo_item != Insumo.TIPO_MATERIA_PRIMA:
                internal_without_source_count += 1
            continue
        current = source_map.setdefault(
            source_recipe_id,
            {
                "id": source_recipe_id,
                "nombre": getattr(getattr(linea, "source_recipe", None), "nombre", ""),
                "line_count": 0,
            },
        )
        if not current["nombre"] and getattr(linea, "source_recipe", None):
            current["nombre"] = linea.source_recipe.nombre
        current["line_count"] += 1

    missing_name_ids = [item["id"] for item in source_map.values() if not item["nombre"]]
    if missing_name_ids:
        for recipe in Receta.objects.filter(id__in=missing_name_ids).only("id", "nombre"):
            if recipe.id in source_map:
                source_map[recipe.id]["nombre"] = recipe.nombre

    derived_parent_snapshot = build_derived_product_upstream_snapshot(receta) if receta is not None else None
    if derived_parent_snapshot:
        source_map.setdefault(
            int(derived_parent_snapshot["parent_recipe_id"]),
            {
                "id": int(derived_parent_snapshot["parent_recipe_id"]),
                "nombre": str(derived_parent_snapshot["parent_recipe_name"]),
                "line_count": 1,
                "derived_parent_link": True,
            },
        )

    upstream_bases = sorted(source_map.values(), key=lambda item: ((item["nombre"] or "").lower(), item["id"]))
    return {
        "upstream_bases": upstream_bases[:8],
        "upstream_base_count": len(upstream_bases),
        "internal_count": internal_count,
        "internal_without_source_count": internal_without_source_count,
        "internal_with_source_count": max(internal_count - internal_without_source_count, 0),
        "empaque_count": empaque_count,
        "mp_count": mp_count,
        "derived_parent_snapshot": derived_parent_snapshot,
    }


def _direct_base_usage_snapshot(lineas: list[LineaReceta]) -> dict[str, object]:
    source_recipe_ids = {
        recipe_id
        for linea in lineas
        if getattr(linea, "insumo_id", None)
        for recipe_id in [_recipe_id_from_derived_code(getattr(getattr(linea, "insumo", None), "codigo", "") or "")]
        if recipe_id
    }
    if not source_recipe_ids:
        return {"count": 0, "base_names": []}

    source_recipe_map = {
        item.id: item
        for item in Receta.objects.filter(id__in=source_recipe_ids).only("id", "nombre", "usa_presentaciones")
    }
    source_recipe_presentaciones = {
        row["receta_id"]: int(row["total"])
        for row in (
            RecetaPresentacion.objects.filter(receta_id__in=source_recipe_ids, activo=True)
            .values("receta_id")
            .annotate(total=Count("id"))
        )
    }

    direct_base_names: list[str] = []
    count = 0
    for linea in lineas:
        insumo = getattr(linea, "insumo", None)
        if not getattr(linea, "insumo_id", None) or not insumo or insumo.tipo_item != Insumo.TIPO_INTERNO:
            continue
        if _derived_code_kind(insumo.codigo or "") != "PREPARACION":
            continue
        source_recipe_id = _recipe_id_from_derived_code(insumo.codigo or "")
        source_recipe = source_recipe_map.get(source_recipe_id) if source_recipe_id else None
        active_presentaciones = int(source_recipe_presentaciones.get(source_recipe_id, 0)) if source_recipe_id else 0
        if source_recipe and source_recipe.usa_presentaciones and active_presentaciones > 0:
            count += 1
            if source_recipe.nombre not in direct_base_names:
                direct_base_names.append(source_recipe.nombre)

    return {
        "count": count,
        "base_names": direct_base_names[:4],
    }


def _recipe_direct_base_snapshot(receta: Receta) -> dict[str, object]:
    cached = getattr(receta, "_direct_base_snapshot_cache", None)
    if cached is not None:
        return cached

    if receta.tipo != Receta.TIPO_PRODUCTO_FINAL:
        snapshot = {
            "count": 0,
            "base_names": [],
            "suggested_count": 0,
            "exact_count": 0,
            "sample_suggestions": [],
        }
        setattr(receta, "_direct_base_snapshot_cache", snapshot)
        return snapshot

    lineas = list(
        receta.lineas.select_related("insumo").only(
            "id",
            "cantidad",
            "insumo_id",
            "insumo__id",
            "insumo__nombre",
            "insumo__codigo",
            "insumo__tipo_item",
        )
    )
    replacement_cache: dict[int, list[dict[str, object]]] = {}
    source_recipe_ids = {
        recipe_id
        for linea in lineas
        if getattr(linea, "insumo_id", None)
        for recipe_id in [_recipe_id_from_derived_code(getattr(getattr(linea, "insumo", None), "codigo", "") or "")]
        if recipe_id
    }
    source_recipe_map = {
        item.id: item
        for item in Receta.objects.filter(id__in=source_recipe_ids).only("id", "nombre", "usa_presentaciones")
    }
    source_recipe_presentaciones = {
        row["receta_id"]: int(row["total"])
        for row in (
            RecetaPresentacion.objects.filter(receta_id__in=source_recipe_ids, activo=True)
            .values("receta_id")
            .annotate(total=Count("id"))
        )
    }

    suggested_count = 0
    exact_count = 0
    sample_suggestions: list[dict[str, object]] = []
    base_names: list[str] = []
    count = 0

    for linea in lineas:
        insumo = getattr(linea, "insumo", None)
        if not getattr(linea, "insumo_id", None) or not insumo or insumo.tipo_item != Insumo.TIPO_INTERNO:
            continue
        if _derived_code_kind(insumo.codigo or "") != "PREPARACION":
            continue
        source_recipe_id = _recipe_id_from_derived_code(insumo.codigo or "")
        source_recipe = source_recipe_map.get(source_recipe_id) if source_recipe_id else None
        active_presentaciones = int(source_recipe_presentaciones.get(source_recipe_id, 0)) if source_recipe_id else 0
        if not source_recipe or not source_recipe.usa_presentaciones or active_presentaciones <= 0:
            continue

        count += 1
        if source_recipe.nombre not in base_names:
            base_names.append(source_recipe.nombre)

        linea.uses_direct_base_in_final = True
        replacement = _suggest_direct_base_replacement(linea, cache=replacement_cache)
        if not replacement:
            continue
        suggested_count += 1
        if replacement["exact_match"]:
            exact_count += 1
        if len(sample_suggestions) < 3:
            sample_suggestions.append(
                {
                    "linea_id": linea.id,
                    "insumo": replacement["insumo"],
                    "presentacion": replacement["presentacion"],
                    "replacement_quantity": replacement["replacement_quantity"],
                    "reason": replacement["reason"],
                    "exact_match": replacement["exact_match"],
                }
            )

    snapshot = {
        "count": count,
        "base_names": base_names[:4],
        "suggested_count": suggested_count,
        "exact_count": exact_count,
        "sample_suggestions": sample_suggestions,
    }
    setattr(receta, "_direct_base_snapshot_cache", snapshot)
    return snapshot


def _active_presentation_derived_candidates(
    receta_base_id: int,
    cache: dict[int, list[dict[str, object]]] | None = None,
) -> list[dict[str, object]]:
    if cache is not None and receta_base_id in cache:
        return cache[receta_base_id]

    presentaciones = list(
        RecetaPresentacion.objects.filter(receta_id=receta_base_id, activo=True).order_by("nombre")
    )
    if not presentaciones:
        if cache is not None:
            cache[receta_base_id] = []
        return []

    code_prefix = f"DERIVADO:RECETA:{receta_base_id}:PRESENTACION:"
    insumos = list(
        Insumo.objects.filter(codigo__startswith=code_prefix, activo=True).only(
            "id",
            "nombre",
            "codigo",
            "unidad_base_id",
        )
    )
    exact_code_map = {item.codigo: item for item in insumos if (item.codigo or "").strip()}
    normalized_name_map = {
        normalizar_nombre(item.nombre or ""): item
        for item in insumos
        if (item.nombre or "").strip()
    }

    candidates: list[dict[str, object]] = []
    for presentacion in presentaciones:
        expected_code = f"DERIVADO:RECETA:{receta_base_id}:PRESENTACION:{presentacion.id}"
        derived_insumo = exact_code_map.get(expected_code)
        if derived_insumo is None:
            derived_insumo = normalized_name_map.get(
                normalizar_nombre(f"{presentacion.receta.nombre} - {presentacion.nombre}")
            )
        if derived_insumo is None:
            continue
        candidates.append(
            {
                "presentacion": presentacion,
                "insumo": derived_insumo,
                "peso": Decimal(str(presentacion.peso_por_unidad_kg or 0)),
            }
        )

    candidates.sort(
        key=lambda item: (
            _presentacion_sort_key(item["presentacion"].nombre),
            item["presentacion"].id,
        )
    )
    if cache is not None:
        cache[receta_base_id] = candidates
    return candidates


def _suggest_direct_base_replacement(
    linea: LineaReceta,
    cache: dict[int, list[dict[str, object]]] | None = None,
) -> dict[str, object] | None:
    if not getattr(linea, "insumo_id", None) or not getattr(linea, "uses_direct_base_in_final", False):
        return None
    source_recipe_id = _recipe_id_from_derived_code(getattr(linea.insumo, "codigo", "") or "")
    if not source_recipe_id:
        return None
    qty = Decimal(str(getattr(linea, "cantidad", None) or 0))
    if qty <= 0:
        return None

    candidates = _active_presentation_derived_candidates(source_recipe_id, cache=cache)
    if not candidates:
        return None

    tolerance = Decimal("0.000001")
    best = min(
        candidates,
        key=lambda item: (
            abs(item["peso"] - qty),
            _presentacion_sort_key(item["presentacion"].nombre),
            item["presentacion"].id,
        ),
    )
    difference = abs(best["peso"] - qty)
    exact_match = difference <= tolerance
    replacement_quantity = Decimal("0")
    if best["peso"] > 0:
        replacement_quantity = (qty / best["peso"]).quantize(Decimal("0.000001"))
    qty_human = f"{qty.quantize(Decimal('0.01'))}"
    peso_human = f"{best['peso'].quantize(Decimal('0.01'))}"
    if exact_match:
        reason = f"Coincide con la cantidad capturada ({qty_human} kg)."
    else:
        reason = f"Es la presentación activa más cercana a {qty_human} kg (objetivo {peso_human} kg)."
    return {
        "source_recipe_id": source_recipe_id,
        "presentacion": best["presentacion"],
        "insumo": best["insumo"],
        "peso": best["peso"],
        "difference": difference,
        "exact_match": exact_match,
        "replacement_quantity": replacement_quantity,
        "reason": reason,
    }


def _recipe_incomplete_erp_item_count(receta: Receta) -> int:
    cached = getattr(receta, "_incomplete_erp_item_count_cache", None)
    if cached is not None:
        return cached
    count = 0
    lineas_qs = receta.lineas.select_related("insumo").only(
        "id",
        "insumo_id",
        "insumo__id",
        "insumo__activo",
        "insumo__tipo_item",
        "insumo__codigo",
        "insumo__categoria",
        "insumo__proveedor_principal_id",
        "insumo__unidad_base_id",
    )
    for linea in lineas_qs:
        if not linea.insumo_id:
            continue
        profile = _insumo_erp_readiness(linea.insumo)
        if not profile["ready"]:
            count += 1
    setattr(receta, "_incomplete_erp_item_count_cache", count)
    return count


def _recipe_master_gap_summary(receta: Receta) -> dict[str, object]:
    cached = getattr(receta, "_master_gap_summary_cache", None)
    if cached is not None:
        return cached

    counters = {
        "unidad": 0,
        "proveedor": 0,
        "categoria": 0,
        "codigo_point": 0,
        "inactivo": 0,
    }
    lineas_qs = receta.lineas.select_related("insumo").only(
        "id",
        "insumo_id",
        "insumo__id",
        "insumo__activo",
        "insumo__tipo_item",
        "insumo__codigo",
        "insumo__codigo_point",
        "insumo__categoria",
        "insumo__proveedor_principal_id",
        "insumo__unidad_base_id",
    )
    for linea in lineas_qs:
        if not linea.insumo_id:
            continue
        profile = _insumo_erp_readiness(linea.insumo)
        missing = set(profile["missing"])
        if "unidad base" in missing:
            counters["unidad"] += 1
        if "proveedor principal" in missing:
            counters["proveedor"] += 1
        if "categoría" in missing:
            counters["categoria"] += 1
        if "inactivo" in missing:
            counters["inactivo"] += 1
        if linea.insumo.activo and not (linea.insumo.codigo_point or "").strip():
            counters["codigo_point"] += 1

    labels = {
        "unidad": "unidad base",
        "proveedor": "proveedor principal",
        "categoria": "categoría",
        "codigo_point": "código comercial",
        "inactivo": "artículo inactivo",
    }
    dominant_key = None
    dominant_count = 0
    for key, count in counters.items():
        if count > dominant_count:
            dominant_key = key
            dominant_count = count

    summary = {
        "counts": counters,
        "dominant_key": dominant_key,
        "dominant_count": dominant_count,
        "dominant_label": labels.get(dominant_key, ""),
    }
    setattr(receta, "_master_gap_summary_cache", summary)
    return summary


def _recipe_master_blockers(
    lineas: list[LineaReceta],
    *,
    receta: Receta | None = None,
    lookback_days: int = 45,
) -> list[dict[str, object]]:
    historico_units = Decimal("0")
    demand_days_count = 0
    if receta and receta.tipo == Receta.TIPO_PRODUCTO_FINAL:
        end_date = timezone.localdate() - timedelta(days=1)
        start_date = end_date - timedelta(days=max(lookback_days - 1, 0))
        historico_qs = VentaHistorica.objects.filter(
            receta_id=receta.id,
            fecha__gte=start_date,
            fecha__lte=end_date,
        )
        historico_units = Decimal(str(historico_qs.aggregate(total=Sum("cantidad"))["total"] or 0))
        demand_days_count = int(historico_qs.values("fecha").distinct().count())

    if historico_units >= Decimal("80") or demand_days_count >= 12:
        priority_label = "Demanda crítica bloqueada"
        priority_tone = "danger"
        priority_detail = "Esta receta ya tiene una señal comercial fuerte y no conviene operar con artículos maestros incompletos."
        is_demand_critical = True
    elif historico_units >= Decimal("30") or demand_days_count >= 5:
        priority_label = "Alta demanda en revisión"
        priority_tone = "warning"
        priority_detail = "La receta ya tiene tracción comercial y conviene cerrar primero estas brechas del maestro."
        is_demand_critical = False
    else:
        priority_label = "Brecha maestra por cerrar"
        priority_tone = "warning"
        priority_detail = "La brecha sigue abierta, aunque todavía sin presión comercial crítica en la ventana observada."
        is_demand_critical = False

    blockers: dict[int, dict[str, object]] = {}
    for linea in lineas:
        if not linea.insumo_id:
            continue
        profile = getattr(linea, "erp_profile", None) or _insumo_erp_readiness(linea.insumo)
        if profile["ready"]:
            continue
        row = blockers.setdefault(
            linea.insumo_id,
            {
                "insumo": linea.insumo,
                "missing": list(profile["missing"]),
                "line_count": 0,
                "line_positions": [],
                "edit_url": reverse("maestros:insumo_update", args=[linea.insumo_id]),
                "historico_units": historico_units.quantize(Decimal("0.1")),
                "demand_days_count": demand_days_count,
                "priority_label": priority_label,
                "priority_tone": priority_tone,
                "priority_detail": priority_detail,
                "is_demand_critical": is_demand_critical,
            },
        )
        row["line_count"] += 1
        row["line_positions"].append(linea.posicion)
    return sorted(
        blockers.values(),
        key=lambda item: (
            0 if item.get("is_demand_critical") else 1,
            -item["line_count"],
            len(item["missing"]),
            (item["insumo"].nombre or "").lower(),
        ),
    )


def _recipe_master_gap_totals_from_blockers(blockers: list[dict[str, object]]) -> dict[str, int]:
    totals = {
        "total": len(blockers),
        "lineas": 0,
        "critical": 0,
        "high_demand": 0,
        "unidad": 0,
        "proveedor": 0,
        "categoria": 0,
        "codigo_point": 0,
        "inactivo": 0,
    }
    for item in blockers:
        missing = set(item.get("missing") or [])
        line_count = int(item.get("line_count") or 0)
        totals["lineas"] += line_count
        if item.get("is_demand_critical"):
            totals["critical"] += 1
        if Decimal(str(item.get("historico_units") or 0)) > 0:
            totals["high_demand"] += 1
        if "unidad base" in missing:
            totals["unidad"] += line_count
        if "proveedor principal" in missing:
            totals["proveedor"] += line_count
        if "categoría" in missing:
            totals["categoria"] += line_count
        if "código comercial" in missing or "código externo" in missing:
            totals["codigo_point"] += line_count
        if "inactivo" in missing:
            totals["inactivo"] += line_count
    return totals


def _recipe_operational_health(receta: Receta) -> dict[str, str]:
    pendientes_attr = getattr(receta, "pendientes_count", None)
    lineas_attr = getattr(receta, "lineas_count", None)
    if pendientes_attr is None:
        pendientes = receta.lineas.filter(
            Q(match_status=LineaReceta.STATUS_NEEDS_REVIEW)
            | Q(match_status=LineaReceta.STATUS_REJECTED)
        ).count()
    else:
        pendientes = int(pendientes_attr or 0)
    if lineas_attr is None:
        lineas = receta.lineas.count()
    else:
        lineas = int(lineas_attr or 0)
    if receta.tipo == Receta.TIPO_PRODUCTO_FINAL:
        if pendientes > 0:
            return {
                "code": "warning",
                "label": "Por validar",
                "description": "Tiene componentes por validar o artículos estándar todavía sin cerrar.",
            }
        if lineas <= 0:
            return {
                "code": "danger",
                "label": "Incompleta",
                "description": "Aún no tiene componentes de armado.",
            }
        lineas_qs = list(
            receta.lineas.select_related("insumo").only(
                "id",
                "insumo_id",
                "insumo__id",
                "insumo__tipo_item",
                "insumo__codigo",
            )
        )
        upstream_snapshot = getattr(receta, "_product_upstream_snapshot_cache", None)
        if upstream_snapshot is None:
            upstream_snapshot = _product_upstream_snapshot(lineas_qs, receta=receta)
            setattr(receta, "_product_upstream_snapshot_cache", upstream_snapshot)
        direct_base_snapshot = _direct_base_usage_snapshot(lineas_qs)
        if direct_base_snapshot["count"] > 0:
            return {
                "code": "warning",
                "label": "Usa base sin presentación",
                "description": "Está consumiendo la base completa aunque la receta origen ya tiene presentaciones activas.",
            }
        if upstream_snapshot["internal_without_source_count"] > 0:
            return {
                "code": "warning",
                "label": "Sin trazabilidad base",
                "description": "Tiene insumos internos ligados que todavía no apuntan a una receta base canónica.",
            }
        if upstream_snapshot["empaque_count"] <= 0:
            return {
                "code": "warning",
                "label": "Sin empaque",
                "description": "Producto final todavía sin empaque ligado en su BOM.",
            }
        incomplete_items = _recipe_incomplete_erp_item_count(receta)
        if incomplete_items > 0:
            return {
                "code": "warning",
                "label": "Maestro incompleto",
                "description": f"Tiene {incomplete_items} artículo(s) ligados que todavía no están listos en el maestro.",
            }
        return {
            "code": "success",
            "label": "Lista para operar",
            "description": "Producto final con estructura lista para costeo, planeación y operación.",
        }

    if not receta.rendimiento_cantidad or not receta.rendimiento_unidad_id:
        return {
            "code": "danger",
            "label": "Sin rendimiento",
            "description": "Falta rendimiento para costeo enterprise.",
        }
    if receta.usa_presentaciones:
        derived_state = _recipe_derived_sync_state(receta)
        presentaciones_activas = int(derived_state["active_presentaciones"])
        if presentaciones_activas <= 0:
            return {
                "code": "warning",
                "label": "Sin derivados",
                "description": "Está marcada con presentaciones pero no tiene derivados activos.",
            }
        if (not derived_state["prep_ready"]) or int(derived_state["derived_presentaciones"]) < presentaciones_activas:
            return {
                "code": "warning",
                "label": "Sincronizar derivados",
                "description": "Tiene presentaciones activas, pero aún faltan derivados maestros o sincronización de base.",
            }
    incomplete_items = _recipe_incomplete_erp_item_count(receta)
    if incomplete_items > 0:
        return {
            "code": "warning",
            "label": "Maestro incompleto",
            "description": f"Tiene {incomplete_items} artículo(s) ligados que todavía no están listos en el maestro.",
        }
    return {
        "code": "success",
        "label": "Lista para operar",
        "description": "Base preparada para costeo, derivados y uso operativo.",
    }


def _module_enablement_handoff_rows(
    receta: Receta,
    module_enablement_cards: list[dict[str, object]],
) -> list[dict[str, object]]:
    fallback_url = reverse("recetas:receta_detail", args=[receta.id])
    meta_map = {
        "Costeo": {
            "owner": "Recetas / Costeo",
            "depends_on": "Maestro + BOM estable",
            "exit_criteria": "Costo trazable y snapshot operativo cerrados.",
        },
        "Costeo base": {
            "owner": "Recetas / Costeo",
            "depends_on": "Rendimiento + unidad",
            "exit_criteria": "La base ya costea con rendimiento y unidad operativa válidos.",
        },
        "Derivados": {
            "owner": "Producción / Recetas",
            "depends_on": "Presentaciones activas",
            "exit_criteria": "Los derivados activos ya existen como artículos internos usables.",
        },
        "Derivados opcionales": {
            "owner": "Producción",
            "depends_on": "Definición operativa",
            "exit_criteria": "La base quedó cerrada para operar sin tamaños derivados.",
        },
        "Producto final": {
            "owner": "Producción / Comercial",
            "depends_on": "Cadena base → derivado",
            "exit_criteria": "La base ya quedó conectada a por lo menos un producto final.",
        },
        "MRP": {
            "owner": "Planeación / Producción",
            "depends_on": "Cadena + costo + maestro",
            "exit_criteria": "La receta ya puede explotar demanda y reabasto sin bloqueos de documento.",
        },
        "MRP / Compras": {
            "owner": "Planeación / Compras",
            "depends_on": "Derivados + consumo final",
            "exit_criteria": "La base ya puede bajar a planeación y abastecimiento aguas abajo.",
        },
        "Compras": {
            "owner": "Compras",
            "depends_on": "Empaque + BOM final",
            "exit_criteria": "Solicitudes y órdenes ya pueden usar este documento sin ambigüedad.",
        },
        "Inventario": {
            "owner": "Inventario / Almacén",
            "depends_on": "Maestro + empaque",
            "exit_criteria": "Inventario ya puede operar el artículo final como referencia estable.",
        },
        "Venta final": {
            "owner": "Operación comercial",
            "depends_on": "Empaque + cadena + costo",
            "exit_criteria": "El artículo ya está listo para operación final consistente.",
        },
        "Uso operativo": {
            "owner": "Producción",
            "depends_on": "Costo + maestro",
            "exit_criteria": "La base ya puede consumirse como insumo interno estable.",
        },
    }
    rows: list[dict[str, object]] = []
    for card in module_enablement_cards:
        label = str(card.get("label") or "Módulo")
        meta = meta_map.get(
            label,
            {
                "owner": "Operación ERP",
                "depends_on": "Cierre documental",
                "exit_criteria": "El documento ya cumple la salida operativa de esta etapa.",
            },
        )
        is_ready = bool(card.get("status"))
        tone = str(card.get("tone") or ("success" if is_ready else "warning"))
        blockers = 0 if is_ready else 1
        rows.append(
            {
                "label": label,
                "owner": meta["owner"],
                "status": "Listo" if is_ready else "Bloqueado",
                "tone": tone,
                "blockers": blockers,
                "completion": 100 if is_ready else 55,
                "depends_on": meta["depends_on"],
                "exit_criteria": meta["exit_criteria"],
                "detail": card.get("detail") or meta["exit_criteria"],
                "next_step": card.get("action_label") or "Abrir documento",
                "url": card.get("action_url") or fallback_url,
                "cta": card.get("action_label") or "Abrir documento",
            }
        )
    return rows


def _matches_recipe_health_filter(receta: Receta, health_filter: str) -> bool:
    health = _recipe_operational_health(receta)
    if health_filter == "listas":
        return health["code"] == "success"
    if health_filter == "pendientes":
        return health["code"] == "warning"
    if health_filter == "incompletas":
        return health["code"] == "danger"
    return True


def _recipe_governance_issues(receta: Receta) -> list[str]:
    issues: list[str] = []
    if not (receta.familia or "").strip():
        issues.append("familia")
    if not (receta.categoria or "").strip():
        issues.append("categoria")
    if _recipe_incomplete_erp_item_count(receta) > 0:
        issues.append("maestro_incompleto")
    if receta.tipo == Receta.TIPO_PREPARACION and (not receta.rendimiento_cantidad or not receta.rendimiento_unidad_id):
        issues.append("rendimiento")
    if receta.tipo == Receta.TIPO_PREPARACION and receta.usa_presentaciones:
        derived_state = _recipe_derived_sync_state(receta)
        presentaciones_activas = int(derived_state["active_presentaciones"])
        if presentaciones_activas <= 0:
            issues.append("derivados")
        elif (not derived_state["prep_ready"]) or int(derived_state["derived_presentaciones"]) < presentaciones_activas:
            issues.append("sync_derivados")
        elif not _recipe_supply_chain_snapshot(receta)["has_downstream_usage"]:
            issues.append("sin_consumo_final")
    if receta.tipo == Receta.TIPO_PRODUCTO_FINAL:
        lineas_count = int(getattr(receta, "lineas_count", 0) or 0)
        if lineas_count <= 0:
            issues.append("componentes")
        else:
            lineas_qs = list(
                receta.lineas.select_related("insumo").only(
                    "id",
                    "insumo_id",
                    "insumo__id",
                    "insumo__tipo_item",
                    "insumo__codigo",
                )
            )
            upstream_snapshot = getattr(receta, "_product_upstream_snapshot_cache", None)
            if upstream_snapshot is None:
                upstream_snapshot = _product_upstream_snapshot(lineas_qs, receta=receta)
                setattr(receta, "_product_upstream_snapshot_cache", upstream_snapshot)
            direct_base_snapshot = getattr(receta, "direct_base_snapshot", None) or _recipe_direct_base_snapshot(receta)
            receta.direct_base_snapshot = direct_base_snapshot
            if int(direct_base_snapshot["count"]) > 0:
                issues.append("base_directa")
            if upstream_snapshot["empaque_count"] <= 0:
                issues.append("sin_empaque")
            if upstream_snapshot["internal_without_source_count"] > 0:
                issues.append("sin_base_origen")
    return issues


def _recipe_primary_action(receta: Receta) -> dict[str, str] | None:
    issues = getattr(receta, "governance_issues", None) or _recipe_governance_issues(receta)
    if "componentes" in issues:
        return {
            "label": "Cargar estructura",
            "url": reverse("recetas:linea_create", args=[receta.id]),
        }
    if "sin_empaque" in issues:
        return {
            "label": "Agregar empaque",
            "url": f"{reverse('recetas:linea_create', args=[receta.id])}?component_kind=EMPAQUE&component_context=empaques",
        }
    if "sin_base_origen" in issues:
        return {
                "label": "Resolver catálogo",
            "url": f"{reverse('recetas:matching_pendientes')}?receta={receta.id}",
        }
    if "base_directa" in issues:
        direct_base_snapshot = getattr(receta, "direct_base_snapshot", None) or _recipe_direct_base_snapshot(receta)
        if int(direct_base_snapshot["suggested_count"] or 0) > 0:
            return {
                "label": "Aplicar derivados sugeridos",
                "url": reverse("recetas:receta_apply_direct_base_replacements", args=[receta.id]),
            }
        return {
            "label": "Corregir base",
            "url": reverse("recetas:receta_detail", args=[receta.id]),
        }
    if "maestro_incompleto" in issues:
        gap_summary = _recipe_master_gap_summary(receta)
        missing_field = gap_summary["dominant_key"] or "unidad"
        return {
            "label": "Revisar maestro",
            "url": (
                f"{reverse('maestros:insumo_list')}?usage_scope=recipes&recipe_scope=finales"
                f"&enterprise_status=incompletos&missing_field={missing_field}&linked_recipe_id={receta.id}"
            ),
        }
    if "derivados" in issues:
        return {
            "label": "Agregar presentación",
            "url": reverse("recetas:presentacion_create", args=[receta.id]),
        }
    if "sync_derivados" in issues:
        return {
            "label": "Sincronizar derivados",
            "url": f"{reverse('recetas:receta_sync_derivados', args=[receta.id])}?next={reverse('recetas:recetas_list')}",
        }
    if "rendimiento" in issues:
        return {
            "label": "Abrir receta",
            "url": reverse("recetas:receta_detail", args=[receta.id]),
        }
    if "sin_consumo_final" in issues:
        return {
            "label": "Crear producto final",
            "url": f"{reverse('recetas:receta_create')}?mode=FINAL&source_base={receta.id}",
        }
    health = getattr(receta, "operational_health", None) or _recipe_operational_health(receta)
    if health["code"] != "success":
        return {
            "label": "Revisar",
            "url": reverse("recetas:receta_detail", args=[receta.id]),
        }
    return None


def _recipe_chain_checkpoints(receta: Receta) -> list[dict[str, str]]:
    checkpoints: list[dict[str, str]] = []

    if receta.tipo == Receta.TIPO_PREPARACION:
        derived_state = _recipe_derived_sync_state(receta)
        supply = getattr(receta, "supply_chain_snapshot", None) or _recipe_supply_chain_snapshot(receta)

        prep_ready = bool(
            receta.rendimiento_cantidad
            and receta.rendimiento_unidad_id
            and derived_state["prep_ready"]
        )
        checkpoints.append(
            {
                "label": "Base",
                "code": "success" if prep_ready else "danger",
                "detail": "Base maestra lista" if prep_ready else "Falta rendimiento o base maestra",
            }
        )

        if receta.usa_presentaciones:
            deriv_ok = (
                int(derived_state["active_presentaciones"]) > 0
                and int(derived_state["derived_presentaciones"]) >= int(derived_state["active_presentaciones"])
            )
            checkpoints.append(
                {
                    "label": "Derivados",
                    "code": "success" if deriv_ok else "warning",
                    "detail": (
                        f"{derived_state['derived_presentaciones']}/{derived_state['active_presentaciones']} activos"
                    ),
                }
            )

        consumo_ok = bool(supply and supply["has_downstream_usage"])
        checkpoints.append(
            {
                "label": "Uso final",
                "code": "success" if consumo_ok else "warning",
                "detail": (
                    f"{supply['downstream_recipe_count']} producto(s) final(es)"
                    if consumo_ok
                    else "Aún sin consumo final"
                ),
            }
        )
    else:
        upstream_snapshot = getattr(receta, "product_upstream_snapshot", None) or getattr(
            receta, "_product_upstream_snapshot_cache", None
        )
        if upstream_snapshot is None:
            lineas_qs = list(
                receta.lineas.select_related("insumo").only(
                    "id",
                    "insumo_id",
                    "insumo__id",
                    "insumo__tipo_item",
                    "insumo__codigo",
                )
            )
            upstream_snapshot = _product_upstream_snapshot(lineas_qs, receta=receta)
            setattr(receta, "_product_upstream_snapshot_cache", upstream_snapshot)

        internal_ok = int(upstream_snapshot["internal_count"]) > 0
        checkpoints.append(
            {
                "label": "Internos",
                "code": "success" if internal_ok else "warning",
                "detail": (
                    f"{upstream_snapshot['internal_count']} ligado(s)"
                    if internal_ok
                    else "Sin internos"
                ),
            }
        )

        trace_ok = (
            int(upstream_snapshot["internal_without_source_count"]) <= 0
            and int(upstream_snapshot["upstream_base_count"]) > 0
        )
        checkpoints.append(
            {
                "label": "Trazabilidad",
                "code": "success" if trace_ok else "warning",
                "detail": (
                    f"{upstream_snapshot['upstream_base_count']} base(s) origen"
                    if trace_ok
                    else "Sin base origen"
                ),
            }
        )

        pack_ok = int(upstream_snapshot["empaque_count"]) > 0
        checkpoints.append(
            {
                "label": "Empaque",
                "code": "success" if pack_ok else "warning",
                "detail": (
                    f"{upstream_snapshot['empaque_count']} ligado(s)"
                    if pack_ok
                    else "Sin empaque"
                ),
            }
        )

    for item in checkpoints:
        item["tone"] = {
            "success": "success",
            "warning": "warning",
            "danger": "danger",
        }.get(item["code"], "secondary")
    return checkpoints


def _recipe_failing_chain_checkpoint_codes(receta: Receta) -> set[str]:
    failing: set[str] = set()
    for item in _recipe_chain_checkpoints(receta):
        if item["code"] == "success":
            continue
        if receta.tipo == Receta.TIPO_PREPARACION:
            if item["label"] == "Base":
                failing.add("base_ready")
            elif item["label"] == "Derivados":
                failing.add("derived_sync")
            elif item["label"] == "Uso final":
                failing.add("final_usage")
        else:
            if item["label"] == "Trazabilidad":
                failing.add("upstream_trace")
            elif item["label"] == "Empaque":
                failing.add("packaging_ready")
            elif item["label"] == "Internos":
                failing.add("internal_components")
    return failing


def _recipe_chain_status(receta: Receta) -> dict[str, str]:
    issues = getattr(receta, "governance_issues", None) or _recipe_governance_issues(receta)
    if receta.tipo == Receta.TIPO_PREPARACION:
        if "rendimiento" in issues:
            return {
                "code": "danger",
                "label": "Base incompleta",
                "description": "Falta rendimiento base para costeo y cadena operativa.",
            }
        if "derivados" in issues:
            return {
                "code": "warning",
                "label": "Sin derivados",
                "description": "La base está marcada para presentaciones, pero aún no tiene derivados activos.",
            }
        if "sync_derivados" in issues:
            return {
                "code": "warning",
                "label": "Sincronizar derivados",
                "description": "Derivados en revisión antes del consumo final.",
            }
        if "sin_consumo_final" in issues:
            return {
                "code": "warning",
                "label": "Sin consumo final",
                "description": "La base tiene derivados, pero todavía no alimenta un producto final.",
            }
        return {
            "code": "success",
            "label": "Cadena lista",
            "description": "La base ya puede alimentar derivados y productos finales.",
        }
    if "componentes" in issues:
        return {
            "code": "danger",
            "label": "Sin estructura",
            "description": "El producto final aún no tiene componentes cargados.",
        }
    if "sin_base_origen" in issues:
        return {
            "code": "warning",
            "label": "Sin trazabilidad de base",
            "description": "Hay internos ligados sin trazabilidad a receta base.",
        }
    if "sin_empaque" in issues:
        return {
            "code": "warning",
            "label": "Sin empaque ligado",
            "description": "Producto final todavía sin empaque ligado en su BOM.",
        }
    return {
        "code": "success",
        "label": "Encadenado",
        "description": "Producto final conectado a base interna y empaque.",
    }


def _recipe_chain_actions_catalog(receta: Receta) -> list[dict[str, str]]:
    actions: list[dict[str, str]] = []
    if receta.tipo == Receta.TIPO_PREPARACION:
        derived_state = getattr(receta, "derived_state", None) or _recipe_derived_sync_state(receta)
        supply = getattr(receta, "supply_chain_snapshot", None) or _recipe_supply_chain_snapshot(receta)
        active_presentaciones = int(derived_state["active_presentaciones"])
        derived_presentaciones = int(derived_state["derived_presentaciones"])
        if receta.usa_presentaciones:
            actions.append(
                {"label": "Presentaciones", "url": reverse("recetas:presentacion_create", args=[receta.id])}
            )
            if active_presentaciones <= 0 or (not derived_state["prep_ready"]) or derived_presentaciones < active_presentaciones:
                actions.append(
                    {
                        "label": "Sincronizar derivados",
                        "url": f"{reverse('recetas:receta_sync_derivados', args=[receta.id])}?next={reverse('recetas:recetas_list')}",
                    }
                )
            if supply and not supply["has_downstream_usage"] and derived_presentaciones > 0:
                actions.append(
                    {
                        "label": "Crear final",
                        "url": f"{reverse('recetas:receta_create')}?mode=FINAL&source_base={receta.id}",
                    }
                )
        elif not receta.rendimiento_cantidad:
            actions.append(
                {"label": "Capturar rendimiento", "url": reverse("recetas:receta_detail", args=[receta.id])}
            )
    else:
        upstream_snapshot = getattr(receta, "product_upstream_snapshot", None) or getattr(
            receta, "_product_upstream_snapshot_cache", None
        )
        if upstream_snapshot is None:
            lineas_qs = list(
                receta.lineas.select_related("insumo").only(
                    "id",
                    "insumo_id",
                    "insumo__id",
                    "insumo__tipo_item",
                    "insumo__codigo",
                )
            )
            upstream_snapshot = _product_upstream_snapshot(lineas_qs, receta=receta)
            setattr(receta, "_product_upstream_snapshot_cache", upstream_snapshot)
        if upstream_snapshot["internal_without_source_count"] > 0:
            actions.append(
                {"label": "Resolver referencias", "url": f"{reverse('recetas:matching_pendientes')}?receta={receta.id}"}
            )
        if upstream_snapshot["empaque_count"] <= 0:
            actions.append(
                {
                    "label": "Agregar empaque",
                    "url": f"{reverse('recetas:linea_create', args=[receta.id])}?component_kind=EMPAQUE&component_context=empaques",
                }
            )
    return actions[:3]


def _recipe_chain_focus_summary(receta: Receta) -> dict[str, Any]:
    chain = getattr(receta, "chain_status_info", None) or _recipe_chain_status(receta)
    issues = getattr(receta, "governance_issues", None) or _recipe_governance_issues(receta)
    primary_action = getattr(receta, "primary_action", None) or _recipe_primary_action(receta)
    direct_base_snapshot = getattr(receta, "direct_base_snapshot", None) or _recipe_direct_base_snapshot(receta)
    supply = getattr(receta, "supply_chain_snapshot", None) or _recipe_supply_chain_snapshot(receta)
    upstream_snapshot = getattr(receta, "product_upstream_snapshot", None) or getattr(
        receta, "_product_upstream_snapshot_cache", None
    )
    if upstream_snapshot is None and receta.tipo == Receta.TIPO_PRODUCTO_FINAL:
        lineas_qs = list(
            receta.lineas.select_related("insumo").only(
                "id",
                "insumo_id",
                "insumo__id",
                "insumo__tipo_item",
                "insumo__codigo",
            )
        )
        upstream_snapshot = _product_upstream_snapshot(lineas_qs, receta=receta)
        setattr(receta, "_product_upstream_snapshot_cache", upstream_snapshot)

    if receta.tipo == Receta.TIPO_PREPARACION:
        if "rendimiento" in issues:
            return {
                "tone": "danger",
                "label": "Capturar rendimiento",
                "summary": "La base todavía no puede costearse ni alimentar derivados porque falta rendimiento total o unidad.",
                "action": primary_action,
            }
        if "derivados" in issues:
            return {
                "tone": "warning",
                "label": "Crear presentaciones",
                "summary": "La base está marcada para derivados, pero aún no tiene presentaciones activas para operar.",
                "action": primary_action,
            }
        if "sync_derivados" in issues:
            derived_state = getattr(receta, "derived_state", None) or _recipe_derived_sync_state(receta)
            return {
                "tone": "warning",
                "label": "Sincronizar derivados",
                "summary": (
                    f"La base tiene {derived_state['active_presentaciones']} presentaciones activas y "
                    f"{derived_state['derived_presentaciones']} derivado(s) operativo(s)."
                ),
                "action": primary_action,
            }
        if "sin_consumo_final" in issues:
            return {
                "tone": "warning",
                "label": "Conectar a producto final",
                "summary": (
                    f"La base ya tiene {supply['presentacion_count']} derivado(s) operativos, "
                    "pero todavía no alimenta ningún producto final."
                ),
                "action": primary_action,
            }
        return {
            "tone": "success",
            "label": "Cadena lista",
            "summary": chain["description"],
            "action": primary_action,
        }

    if "componentes" in issues:
        return {
            "tone": "danger",
            "label": "Completar estructura",
            "summary": "El producto final aún no tiene componentes internos cargados para costeo y producción.",
            "action": primary_action,
        }
    if "base_directa" in issues:
        suggested_count = int(direct_base_snapshot.get("suggested_count") or 0)
        return {
            "tone": "warning",
            "label": "Resolver derivados",
            "summary": (
                f"El producto final consume base directa. Hay {suggested_count} reemplazo(s) sugerido(s) listo(s) para aplicar."
                if suggested_count
                else "El producto final consume base directa y aún requiere corrección manual hacia un derivado operativo."
            ),
            "action": primary_action,
        }
    if "sin_base_origen" in issues:
        missing_count = int(upstream_snapshot["internal_without_source_count"] or 0) if upstream_snapshot else 0
        return {
            "tone": "warning",
            "label": "Resolver referencias",
            "summary": f"Hay {missing_count} interno(s) sin trazabilidad a receta base dentro del producto final.",
            "action": primary_action,
        }
    if "sin_empaque" in issues:
        return {
            "tone": "warning",
            "label": "Agregar empaque",
            "summary": "El producto final aún no tiene empaque ligado, por lo que la cadena comercial sigue incompleta.",
            "action": primary_action,
        }
    return {
        "tone": "success",
        "label": "Cadena lista",
        "summary": chain["description"],
        "action": primary_action,
    }


def _recipe_catalog_governance_rows(
    *,
    total_pendientes: int,
    health_summary: dict[str, int],
    master_gap_totals: dict[str, int],
    chain_focus: dict[str, Any],
) -> list[dict[str, object]]:
    master_total = int(master_gap_totals.get("total") or 0)
    incompletas = int(health_summary.get("incompletas") or 0)
    action = chain_focus.get("action") or {}
    return [
        {
            "front": "Catálogo BOM",
            "owner": "Producción / Recetas",
            "blockers": total_pendientes,
            "completion": 100 if total_pendientes == 0 else 70,
            "detail": (
                "El catálogo ya puede operar sin documentos abiertos."
                if total_pendientes == 0
                else f"{total_pendientes} receta(s) siguen en revisión dentro del catálogo."
            ),
            "next_step": (
                "Mantener monitoreo preventivo."
                if total_pendientes == 0
                else "Cerrar brechas estructurales y documentales."
            ),
            "url": reverse("recetas:recetas_list"),
            "cta": "Abrir catálogo",
        },
        {
            "front": "Maestro de artículos",
            "owner": "Maestros / Inventario",
            "blockers": master_total,
            "completion": 100 if master_total == 0 else 75,
            "detail": (
                "El maestro ya no bloquea recetas activas."
                if master_total == 0
                else f"{master_total} artículo(s) ligados siguen incompletos para operar."
            ),
            "next_step": (
                "Mantener integridad maestra."
                if master_total == 0
                else "Completar el maestro antes de seguir liberando documentos."
            ),
            "url": reverse("maestros:insumo_list"),
            "cta": "Abrir maestro",
        },
        {
            "front": "Cadena operativa",
            "owner": "ERP / Operaciones",
            "blockers": incompletas,
            "completion": 100 if incompletas == 0 else 85,
            "detail": chain_focus.get("summary") or "La cadena base → derivado → producto final ya está bajo control.",
            "next_step": (
                chain_focus.get("action_detail")
                or ("Mantener la cadena cerrada." if incompletas == 0 else "Resolver el bloqueo dominante de cadena.")
            ),
            "url": action.get("url") or reverse("recetas:recetas_list"),
            "cta": action.get("label") or "Abrir siguiente paso",
        },
    ]


def _recipe_detail_governance_rows(
    *,
    receta: Receta,
    recipe_master_gap_totals: dict[str, int],
    total_revision: int,
    total_sin_match: int,
    release_gate_progress: dict[str, int],
    chain_focus_summary: dict[str, Any],
) -> list[dict[str, object]]:
    action = chain_focus_summary.get("action") or {}
    master_total = int(recipe_master_gap_totals.get("total") or 0)
    return [
        {
            "front": "Documento BOM",
            "owner": "Producción / Recetas",
            "blockers": total_revision + total_sin_match,
            "completion": int(release_gate_progress.get("pct") or 0),
            "detail": (
                "El documento ya está listo para costeo y operación."
                if (total_revision + total_sin_match) == 0
                else f"{total_revision + total_sin_match} componente(s) siguen abiertos en revisión documental."
            ),
            "next_step": (
                "Mantener monitoreo del documento."
                if (total_revision + total_sin_match) == 0
                else "Cerrar componentes pendientes y referencias abiertas."
            ),
            "url": reverse("recetas:receta_detail", args=[receta.id]),
            "cta": "Abrir documento",
        },
        {
            "front": "Maestro de artículos",
            "owner": "Maestros / Inventario",
            "blockers": master_total,
            "completion": 100 if master_total == 0 else 75,
            "detail": (
                "Los artículos ligados ya están listos para operar."
                if master_total == 0
                else f"{master_total} artículo(s) ligados siguen incompletos en el maestro."
            ),
            "next_step": (
                "Mantener integridad maestra."
                if master_total == 0
                else "Completar datos maestros de los artículos bloqueados."
            ),
            "url": reverse("maestros:insumo_list"),
            "cta": "Abrir maestro",
        },
        {
            "front": "Cadena ERP",
            "owner": "ERP / Operaciones",
            "blockers": 0 if int(release_gate_progress.get("pct") or 0) >= 100 else 1,
            "completion": int(release_gate_progress.get("pct") or 0),
            "detail": chain_focus_summary.get("summary") or "La cadena operativa del documento está bajo control.",
            "next_step": chain_focus_summary.get("action_detail") or "Completar la siguiente acción recomendada.",
            "url": action.get("url") or reverse("recetas:receta_detail", args=[receta.id]),
            "cta": action.get("label") or "Abrir siguiente paso",
        },
    ]


def _recipe_detail_trunk_handoff_rows(
    *,
    receta: Receta,
    recipe_master_gap_totals: dict[str, int],
    total_revision: int,
    total_sin_match: int,
    release_gate_progress: dict[str, int],
    chain_focus_summary: dict[str, Any],
    is_producto_final: bool,
    is_base_con_presentaciones: bool,
    governance_issues: list[str],
    supply_chain_snapshot: dict[str, object] | None,
    product_upstream_snapshot: dict[str, object] | None,
) -> list[dict[str, object]]:
    master_total = int(recipe_master_gap_totals.get("total") or 0)
    document_blockers = total_revision + total_sin_match
    chain_action = chain_focus_summary.get("action") or {}
    chain_blockers = 0 if int(release_gate_progress.get("pct") or 0) >= 100 else 1
    compras_blockers = master_total + chain_blockers
    inventory_blockers = master_total

    if is_producto_final and "sin_empaque" in governance_issues:
        compras_blockers += 1
        inventory_blockers += 1
    if is_base_con_presentaciones and supply_chain_snapshot and not bool(supply_chain_snapshot.get("has_downstream_usage")):
        compras_blockers += 1
    if is_producto_final and product_upstream_snapshot:
        if int(product_upstream_snapshot.get("upstream_base_count") or 0) <= 0:
            compras_blockers += 1

    recetas_ready = (document_blockers + master_total + chain_blockers) == 0
    compras_ready = compras_blockers == 0
    inventario_ready = inventory_blockers == 0

    return [
        {
            "label": "Recetas / BOM",
            "owner": "Producción / Recetas",
            "status": "Listo para operar" if recetas_ready else "Bloqueado",
            "tone": "success" if recetas_ready else "warning",
            "blockers": document_blockers + master_total + chain_blockers,
            "completion": 100 if recetas_ready else max(0, min(95, int(release_gate_progress.get("pct") or 0))),
            "depends_on": "Documento limpio + catálogo ERP + cadena cerrada",
            "exit_criteria": "La receta debe costear, normalizar y operar sin brechas documentales ni artículos abiertos.",
            "detail": (
                "La receta ya puede operar como documento BOM estable."
                if recetas_ready
                else "Todavía hay brechas en documento, catálogo ERP o cadena operativa."
            ),
            "next_step": chain_focus_summary.get("action_detail") or "Cerrar brechas del documento",
            "url": chain_action.get("url") or reverse("recetas:receta_detail", args=[receta.id]),
            "cta": chain_action.get("label") or "Abrir documento",
        },
        {
            "label": "Compras documentales",
            "owner": "Compras / Planeación",
            "status": "Listo para operar" if compras_ready else "Bloqueado",
            "tone": "success" if compras_ready else "warning",
            "blockers": compras_blockers,
            "completion": 100 if compras_ready else max(0, 100 - (compras_blockers * 12)),
            "depends_on": "Producto operable + empaques cerrados + maestro consistente",
            "exit_criteria": "Solicitudes, órdenes y recepciones deben tomar esta receta sin ambigüedad estructural.",
            "detail": (
                "Compras ya puede trabajar este documento como referencia estable."
                if compras_ready
                else "Compras sigue condicionado por maestro, empaque o cadena operativa."
            ),
            "next_step": "Abrir compras" if compras_ready else "Cerrar empaque y catálogo ERP",
            "url": reverse("compras:solicitudes") if compras_ready else reverse("maestros:insumo_list"),
            "cta": "Abrir compras" if compras_ready else "Abrir catálogo ERP",
        },
        {
            "label": "Inventario / Reabasto",
            "owner": "Inventario / Almacén",
            "status": "Listo para operar" if inventario_ready else "Bloqueado",
            "tone": "success" if inventario_ready else "warning",
            "blockers": inventory_blockers,
            "completion": 100 if inventario_ready else max(0, 100 - (inventory_blockers * 15)),
            "depends_on": "Artículos listos + empaque operativo + salida documental estable",
            "exit_criteria": "Inventario debe sostener existencia, stock y reabasto sin artículos abiertos del maestro.",
            "detail": (
                "Inventario ya puede operar esta receta como referencia estable."
                if inventario_ready
                else "Inventario todavía requiere cierre del catálogo ERP o empaque para operar estable."
            ),
            "next_step": "Abrir inventario" if inventario_ready else "Cerrar catálogo ERP",
            "url": reverse("inventario:existencias") if inventario_ready else reverse("maestros:insumo_list"),
            "cta": "Abrir inventario" if inventario_ready else "Abrir catálogo ERP",
        },
    ]


def _recipe_enterprise_stage(receta: Receta) -> dict[str, str]:
    issues = getattr(receta, "governance_issues", None) or _recipe_governance_issues(receta)
    chain = getattr(receta, "chain_status_info", None) or _recipe_chain_status(receta)
    supply = getattr(receta, "supply_chain_snapshot", None) or _recipe_supply_chain_snapshot(receta)
    upstream_snapshot = getattr(receta, "product_upstream_snapshot", None) or getattr(
        receta, "_product_upstream_snapshot_cache", None
    )

    if receta.tipo == Receta.TIPO_PREPARACION:
        if "rendimiento" in issues:
            return {
                "code": "base_setup",
                "label": "Base en configuración",
                "tone": "danger",
                "summary": "Falta rendimiento o unidad para cerrar la batida base.",
            }
        if receta.usa_presentaciones:
            if "derivados" in issues or "sync_derivados" in issues:
                return {
                    "code": "derivados_setup",
                    "label": "Derivados en configuración",
                    "tone": "warning",
                    "summary": "La base ya existe, pero sus presentaciones o derivados aún no están listos.",
                }
            if "sin_consumo_final" in issues:
                return {
                    "code": "ready_for_final",
                    "label": "Base liberada a final",
                    "tone": "warning",
                    "summary": "La base y sus derivados ya están listos, pero todavía no alimentan un producto final.",
                }
            return {
                "code": "feeding_final",
                "label": "Alimentando producto final",
                "tone": "success",
                "summary": "La base ya opera con derivados activos y consumo final.",
            }
        return {
            "code": "base_operativa",
            "label": "Base operativa",
            "tone": "success" if chain["code"] == "success" else "warning",
            "summary": "La receta base ya puede costearse y operar como insumo interno.",
        }

    if "componentes" in issues:
        return {
            "code": "final_setup",
            "label": "Armado inicial",
            "tone": "danger",
            "summary": "El producto final aún no tiene estructura operativa suficiente.",
        }
    if "base_directa" in issues or "sin_base_origen" in issues:
        return {
            "code": "final_normalization",
            "label": "Ajuste de estructura",
            "tone": "warning",
            "summary": "El producto final requiere ajustar bases, derivados u origen de componentes.",
        }
    if "sin_empaque" in issues:
        return {
            "code": "final_packaging",
            "label": "Cierre comercial",
            "tone": "warning",
            "summary": "La estructura productiva está casi completa, pero falta empaque final para cerrar salida comercial.",
        }
    if upstream_snapshot and int(upstream_snapshot.get("upstream_base_count") or 0) > 0:
        return {
            "code": "final_ready",
            "label": "Listo para operación ERP",
            "tone": "success",
            "summary": "El producto final ya está conectado a su cadena base y listo para costeo/abasto.",
        }
    return {
        "code": "final_defined",
        "label": "Producto definido",
        "tone": "success" if chain["code"] == "success" else "warning",
        "summary": "El producto final está estructurado y sin bloqueos críticos visibles.",
    }


def _matches_recipe_enterprise_stage(receta: Receta, selected_stage: str) -> bool:
    return _recipe_enterprise_stage(receta)["code"] == selected_stage


def _recipe_enterprise_stage_playbook(receta: Receta) -> list[dict[str, Any]]:
    stage = _recipe_enterprise_stage(receta)
    checkpoints = _recipe_chain_checkpoints(receta)
    checkpoint_map = {item["label"]: item for item in checkpoints}
    items: list[dict[str, Any]] = []

    if receta.tipo == Receta.TIPO_PREPARACION:
        derived_state = _recipe_derived_sync_state(receta)
        supply = getattr(receta, "supply_chain_snapshot", None) or _recipe_supply_chain_snapshot(receta)
        base_ok = bool(receta.rendimiento_cantidad and receta.rendimiento_unidad_id and derived_state["prep_ready"])
        items.append(
            {
                "label": "Base y rendimiento",
                "done": base_ok,
                "detail": checkpoint_map.get("Base", {}).get("detail", "Sin validar"),
                "action_label": None if base_ok else "Abrir receta",
                "action_url": None if base_ok else reverse("recetas:receta_detail", args=[receta.id]),
            }
        )
        if receta.usa_presentaciones:
            derivados_ok = (
                int(derived_state["active_presentaciones"]) > 0
                and int(derived_state["derived_presentaciones"]) >= int(derived_state["active_presentaciones"])
            )
            items.append(
                {
                    "label": "Presentaciones y derivados",
                    "done": derivados_ok,
                    "detail": checkpoint_map.get("Derivados", {}).get("detail", "Sin validar"),
                    "action_label": None if derivados_ok else "Sincronizar derivados",
                    "action_url": None if derivados_ok else reverse("recetas:receta_sync_derivados", args=[receta.id]),
                }
            )
        uso_final_ok = bool(supply and supply["has_downstream_usage"])
        items.append(
            {
                "label": "Consumo en producto final",
                "done": uso_final_ok,
                "detail": checkpoint_map.get("Uso final", {}).get("detail", "Sin validar"),
                "action_label": None if uso_final_ok else "Crear producto final",
                "action_url": None if uso_final_ok else f"{reverse('recetas:receta_create')}?mode=FINAL&source_base={receta.id}",
            }
        )
    else:
        upstream_snapshot = getattr(receta, "product_upstream_snapshot", None) or getattr(
            receta,
            "_product_upstream_snapshot_cache",
            None,
        )
        if upstream_snapshot is None:
            lineas_qs = list(
                receta.lineas.select_related("insumo").only(
                    "id",
                    "insumo_id",
                    "insumo__id",
                    "insumo__tipo_item",
                    "insumo__codigo",
                )
            )
            upstream_snapshot = _product_upstream_snapshot(lineas_qs, receta=receta)
            setattr(receta, "_product_upstream_snapshot_cache", upstream_snapshot)
        internos_ok = int(upstream_snapshot["internal_count"]) > 0
        items.append(
            {
                "label": "Insumos internos",
                "done": internos_ok,
                "detail": checkpoint_map.get("Internos", {}).get("detail", "Sin validar"),
                "action_label": None if internos_ok else "Agregar interno",
                "action_url": None if internos_ok else f"{reverse('recetas:linea_create', args=[receta.id])}?component_kind=INSUMO_INTERNO&component_context=internos",
            }
        )
        trazabilidad_ok = (
            int(upstream_snapshot["internal_without_source_count"]) <= 0
            and int(upstream_snapshot["upstream_base_count"]) > 0
        )
        items.append(
            {
                "label": "Trazabilidad a base",
                "done": trazabilidad_ok,
                "detail": checkpoint_map.get("Trazabilidad", {}).get("detail", "Sin validar"),
                "action_label": None if trazabilidad_ok else "Revisar estructura",
                "action_url": None if trazabilidad_ok else reverse("recetas:receta_detail", args=[receta.id]),
            }
        )
        empaque_ok = int(upstream_snapshot["empaque_count"]) > 0
        items.append(
            {
                "label": "Empaque final",
                "done": empaque_ok,
                "detail": checkpoint_map.get("Empaque", {}).get("detail", "Sin validar"),
                "action_label": None if empaque_ok else "Agregar empaque",
                "action_url": None if empaque_ok else f"{reverse('recetas:linea_create', args=[receta.id])}?component_kind=EMPAQUE&component_context=empaques",
            }
        )

    summary_map = {
        "base_setup": "Completa la base mínima para que el costo y el flujo queden listos.",
        "derivados_setup": "Cierra derivados y presentaciones para habilitar consumo final.",
        "ready_for_final": "La base ya puede alimentar productos finales sin bloqueo.",
        "feeding_final": "La base ya está conectada a productos finales operativos.",
        "base_operativa": "La base simple ya puede costearse y operar como insumo interno.",
        "final_setup": "Completa el rompecabezas del producto final antes de costear o planear.",
        "final_normalization": "Valida artículos y reemplaza bases directas para estabilizar la estructura.",
        "final_packaging": "Agrega empaque final para cerrar el producto en operación.",
        "final_ready": "El producto final está completo para costeo, MRP y compras.",
        "final_defined": "El producto final ya está definido y estable en el catálogo.",
    }
    for item in items:
        item["tone"] = "success" if item["done"] else "warning"
    return [{"label": "Resumen de etapa", "done": stage["tone"] == "success", "detail": summary_map.get(stage["code"], stage["summary"]), "action_label": None, "action_url": None}] + items


def _recipe_stage_progress(playbook: list[dict[str, Any]]) -> dict[str, int]:
    actionable = [item for item in (playbook or []) if item.get("label") != "Resumen de etapa"]
    total = len(actionable)
    completed = sum(1 for item in actionable if item.get("done"))
    pct = int(round((completed / total) * 100)) if total else 100
    return {
        "completed": completed,
        "total": total,
        "pct": pct,
    }


def _recipe_document_status(receta: Receta) -> dict[str, str]:
    stage = getattr(receta, "enterprise_stage", None) or _recipe_enterprise_stage(receta)
    health = getattr(receta, "operational_health", None) or _recipe_operational_health(receta)
    chain = getattr(receta, "chain_status_info", None) or _recipe_chain_status(receta)
    progress = getattr(receta, "enterprise_stage_progress", None) or _recipe_stage_progress(
        getattr(receta, "enterprise_stage_playbook", None) or _recipe_enterprise_stage_playbook(receta)
    )
    lineas_count = int(getattr(receta, "lineas_count", 0) or 0)

    if stage["tone"] == "danger" or health["code"] == "danger":
        return {
            "code": "blocked",
            "label": "Bloqueado",
            "tone": "danger",
            "detail": "Todavía no cumple las reglas mínimas para operar en el ERP.",
        }
    if lineas_count <= 0:
        return {
            "code": "draft",
            "label": "Borrador",
            "tone": "warning",
            "detail": "Aún no tiene estructura suficiente para validación operativa.",
        }
    if progress["pct"] >= 100 and health["code"] == "success" and chain["code"] == "success":
        return {
            "code": "operable",
            "label": "Operable",
            "tone": "success",
            "detail": "Documento listo para costeo, MRP y abastecimiento.",
        }
    return {
        "code": "validation",
        "label": "Por validar",
        "tone": "warning",
        "detail": "El documento ya arrancó, pero aún requiere cierre operativo.",
    }


def _build_recipe_chain_focus(
    *,
    vista: str,
    chain_summary: dict[str, int],
    checkpoint_summary: dict[str, int],
    governance_summary: dict[str, int],
    filters: dict[str, str],
) -> dict[str, str]:
    base_query = {
        key: value
        for key, value in filters.items()
        if value and key not in {"chain_status", "chain_checkpoint", "governance_issue"}
    }

    def _list_url(**extra: str) -> str:
        query = {**base_query, **{k: v for k, v in extra.items() if v}}
        encoded = urlencode(query)
        return f"{reverse('recetas:recetas_list')}?{encoded}" if encoded else reverse("recetas:recetas_list")

    if vista == "productos":
        priorities = [
            (
                "internal_components",
                max(
                    int(checkpoint_summary["internal_components"] or 0),
                    int(governance_summary.get("componentes") or 0),
                ),
                {
                    "label": "Internos faltantes",
                    "tone": "danger",
                    "summary": "Hay productos finales sin componentes internos ligados en la estructura.",
                    "action_label": "Cargar internos",
                    "action_detail": "Completa la estructura base del producto final antes de costear, producir o abastecer.",
                },
            ),
            (
                "upstream_trace",
                max(
                    int(checkpoint_summary["upstream_trace"] or 0),
                    int(governance_summary.get("sin_base_origen") or 0),
                ),
                {
                    "label": "Trazabilidad en revisión",
                    "tone": "warning",
                    "summary": "Hay productos finales con insumos internos todavía sin base origen identificada.",
                    "action_label": "Resolver referencias",
                    "action_detail": "Resuelve la referencia estándar para ligar cada interno a su batida o receta base.",
                },
            ),
            (
                "packaging_ready",
                max(
                    int(checkpoint_summary["packaging_ready"] or 0),
                    int(governance_summary.get("sin_empaque") or 0),
                ),
                {
                    "label": "Empaque faltante",
                    "tone": "warning",
                    "summary": "Hay productos finales sin empaque ligado, por lo que el costo y la salida no están cerrados.",
                    "action_label": "Agregar empaques",
                    "action_detail": "Completa el empaque final de la estructura para cerrar costeo, surtido y venta.",
                },
            ),
        ]
    elif vista in {"subinsumos", "insumos"}:
        priorities = [
            (
                "base_ready",
                max(
                    int(checkpoint_summary["base_ready"] or 0),
                    int(governance_summary.get("rendimiento") or 0),
                ),
                {
                    "label": "Base por cerrar",
                    "tone": "danger",
                    "summary": "Hay batidas o mezclas base sin rendimiento o sin estructura mínima para operar.",
                    "action_label": "Completar bases",
                    "action_detail": "Captura rendimiento, unidad y base operativa antes de derivar o costear.",
                },
            ),
            (
                "derived_sync",
                max(
                    int(checkpoint_summary["derived_sync"] or 0),
                    int(governance_summary.get("sync_derivados") or 0),
                    int(governance_summary.get("derivados") or 0),
                ),
                {
                    "label": "Sincronización de derivados",
                    "tone": "warning",
                    "summary": "Hay recetas con presentaciones activas que aún no tienen derivados operativos actualizados.",
                    "action_label": "Sincronizar derivados",
                    "action_detail": "Genera o actualiza subinsumos derivados para que producción y costeo consuman artículos correctos.",
                },
            ),
            (
                "final_usage",
                max(
                    int(checkpoint_summary["final_usage"] or 0),
                    int(governance_summary.get("sin_consumo_final") or 0),
                ),
                {
                    "label": "Sin consumo final",
                    "tone": "warning",
                    "summary": "Hay bases ya derivadas que todavía no alimentan ningún producto final.",
                    "action_label": "Crear productos finales",
                    "action_detail": "Conecta la base con sus productos de venta para cerrar la cadena operativa.",
                },
            ),
        ]
    else:
        priorities = [
            (
                "base_ready",
                max(
                    int(checkpoint_summary["base_ready"] or 0),
                    int(governance_summary.get("rendimiento") or 0),
                ),
                {
                    "label": "Base por validar",
                    "tone": "danger",
                    "summary": "Hay recetas base incompletas que están frenando la cadena operativa.",
                    "action_label": "Completar bases",
                    "action_detail": "Corrige rendimiento y estructura base antes de derivar, consumir o costear.",
                },
            ),
            (
                "internal_components",
                max(
                    int(checkpoint_summary["internal_components"] or 0),
                    int(governance_summary.get("componentes") or 0),
                ),
                {
                    "label": "Internos faltantes",
                    "tone": "danger",
                    "summary": "Hay productos finales sin insumos internos suficientes para operar como BOM enterprise.",
                    "action_label": "Cargar internos",
                    "action_detail": "Completa los insumos internos para evitar productos finales incompletos.",
                },
            ),
            (
                "derived_sync",
                max(
                    int(checkpoint_summary["derived_sync"] or 0),
                    int(governance_summary.get("sync_derivados") or 0),
                    int(governance_summary.get("derivados") or 0),
                ),
                {
                    "label": "Sincronización de derivados",
                    "tone": "warning",
                    "summary": "Hay bases con presentaciones activas que aún no han actualizado sus derivados operativos.",
                    "action_label": "Sincronizar derivados",
                    "action_detail": "Alinea bases y subinsumos antes de continuar con armado final.",
                },
            ),
            (
                "packaging_ready",
                max(
                    int(checkpoint_summary["packaging_ready"] or 0),
                    int(governance_summary.get("sin_empaque") or 0),
                ),
                {
                    "label": "Empaque faltante",
                    "tone": "warning",
                    "summary": "Hay productos finales sin empaque y no deberían considerarse listos para salida.",
                    "action_label": "Agregar empaques",
                    "action_detail": "Completa el empaque final para cerrar costo unitario y operación de salida.",
                },
            ),
            (
                "upstream_trace",
                checkpoint_summary["upstream_trace"],
                {
                    "label": "Trazabilidad por validar",
                    "tone": "warning",
                    "summary": "Hay productos finales cuya estructura interna no tiene trazabilidad cerrada a una base origen.",
                    "action_label": "Resolver referencias",
                    "action_detail": "Corrige integración y catálogo estándar para estabilizar costeo, MRP y compras.",
                },
            ),
            (
                "final_usage",
                checkpoint_summary["final_usage"],
                {
                    "label": "Sin consumo final",
                    "tone": "warning",
                    "summary": "Hay bases derivadas que todavía no se consumen en un producto final.",
                    "action_label": "Crear productos finales",
                    "action_detail": "Conecta la base a un producto final para cerrar el flujo enterprise.",
                },
            ),
        ]

    for checkpoint_code, count, meta in priorities:
        if int(count or 0) > 0:
            return {
                "checkpoint": checkpoint_code,
                "count": str(count),
                "label": meta["label"],
                "tone": meta["tone"],
                "summary": meta["summary"],
                "action_label": meta["action_label"],
                "action_url": _list_url(chain_checkpoint=checkpoint_code),
                "action_detail": meta["action_detail"],
            }

    return {
        "checkpoint": "",
        "count": str(chain_summary["listas"]),
        "label": "Cadena operativa lista",
        "tone": "success",
        "summary": "Las recetas visibles ya cumplen la cadena enterprise esperada para su etapa operativa.",
        "action_label": "Abrir recetas listas",
        "action_url": _list_url(chain_status="listas"),
        "action_detail": "Revisa recetas listas para costeo, producción o armado final sin bloqueos de cadena.",
    }


def _reabasto_enterprise_board(
    *,
    recetas_producto: list[Receta],
    inventario_map: dict[int, Any],
    fecha_operacion: date,
) -> dict[str, Any]:
    blocker_cards_map = {
        "sin_inventario": {"key": "sin_inventario", "label": "Sin inventario CEDIS", "tone": "danger", "count": 0},
        "sin_empaque": {"key": "sin_empaque", "label": "Sin empaque", "tone": "warning", "count": 0},
        "maestro_incompleto": {"key": "maestro_incompleto", "label": "Maestro incompleto", "tone": "warning", "count": 0},
        "pendiente": {"key": "pendiente", "label": "Receta por validar", "tone": "danger", "count": 0},
        "lista": {"key": "lista", "label": "Lista para operar", "tone": "success", "count": 0},
    }
    detail_rows: list[dict[str, Any]] = []
    current_url = f"{reverse('recetas:reabasto_cedis')}?fecha={fecha_operacion.isoformat()}"
    blocked_recipe_ids: set[int] = set()
    for receta in recetas_producto:
        health = _recipe_operational_health(receta)
        blockers: list[dict[str, str]] = []
        if health["label"] == "Sin empaque":
            blockers.append(
                {
                    "key": "sin_empaque",
                    "label": "Sin empaque",
                    "detail": health["description"],
                    "action_label": "Agregar empaque",
                    "action_url": f"{reverse('recetas:linea_create', args=[receta.id])}?component_kind=EMPAQUE&component_context=empaques",
                }
            )
        elif health["label"] == "Maestro incompleto":
            blockers.append(
                {
                    "key": "maestro_incompleto",
                    "label": "Maestro incompleto",
                    "detail": health["description"],
                    "action_label": "Abrir receta",
                    "action_url": reverse("recetas:receta_detail", args=[receta.id]),
                }
            )
        elif health["code"] != "success":
            blockers.append(
                {
                    "key": "pendiente",
                    "label": "Receta por validar",
                    "detail": health["description"],
                    "action_label": "Abrir receta",
                    "action_url": reverse("recetas:receta_detail", args=[receta.id]),
                }
            )
        if receta.id not in inventario_map:
            blockers.append(
                {
                    "key": "sin_inventario",
                    "label": "Sin inventario CEDIS",
                    "detail": "El producto no tiene stock cargado en inventario CEDIS para abastecimiento.",
                    "action_label": "Registrar inventario",
                    "action_url": f"{current_url}#inventario-cedis",
                }
            )

        if blockers:
            blocked_recipe_ids.add(receta.id)
            seen_blocker_keys: set[str] = set()
            for blocker in blockers:
                if blocker["key"] in seen_blocker_keys:
                    continue
                blocker_cards_map[blocker["key"]]["count"] += 1
                seen_blocker_keys.add(blocker["key"])
            for index, blocker in enumerate(blockers):
                detail_rows.append(
                    {
                        "blocker_key": blocker["key"],
                        "receta_id": receta.id,
                        "receta_nombre": receta.nombre,
                        "blocker_label": blocker["label"],
                        "detail": blocker["detail"],
                        "action_label": blocker["action_label"],
                        "action_url": blocker["action_url"],
                        "is_primary": index == 0,
                    }
                )
        else:
            blocker_cards_map["lista"]["count"] += 1

    blocker_cards = [
        item for item in blocker_cards_map.values() if item["count"] > 0 or item["label"] == "Listas"
    ]
    blocker_cards.sort(key=lambda item: (item["label"] == "Listas", -int(item["count"]), item["label"]))
    detail_rows.sort(key=lambda item: (item["blocker_label"], item["receta_nombre"]))
    return {
        "blocker_cards": blocker_cards,
        "detail_rows": detail_rows[:12],
        "blocked_total": len(blocked_recipe_ids),
        "ready_total": blocker_cards_map["lista"]["count"],
    }


def _find_reabasto_plan(fecha_operacion: date) -> PlanProduccion | None:
    marker = f"[AUTO_REABASTO_CEDIS:{fecha_operacion.isoformat()}]"
    nombre_plan = f"CEDIS Reabasto {fecha_operacion.isoformat()}"
    return (
        PlanProduccion.objects.filter(
            fecha_produccion=fecha_operacion,
        )
        .filter(Q(nombre=nombre_plan) | Q(notas__icontains=marker))
        .order_by("-id")
        .first()
    )


def _reabasto_enterprise_context(fecha_operacion: date) -> dict[str, Any]:
    recetas_producto = list(
        Receta.objects.filter(tipo=Receta.TIPO_PRODUCTO_FINAL)
        .order_by("nombre")
        .only("id", "nombre", "codigo_point")
    )
    if not recetas_producto:
        recetas_producto = list(Receta.objects.order_by("nombre").only("id", "nombre", "codigo_point")[:400])

    inventario_cedis = list(
        InventarioCedisProducto.objects.select_related("receta")
        .filter(receta__in=[r.id for r in recetas_producto])
        .order_by("receta__nombre")
    )
    inventario_map = {inv.receta_id: inv for inv in inventario_cedis}
    productos_sin_inventario = [r for r in recetas_producto if r.id not in inventario_map]
    reabasto_enterprise_board = _reabasto_enterprise_board(
        recetas_producto=recetas_producto,
        inventario_map=inventario_map,
        fecha_operacion=fecha_operacion,
    )
    return {
        "recetas_producto": recetas_producto,
        "inventario_cedis": inventario_cedis,
        "inventario_map": inventario_map,
        "productos_sin_inventario": productos_sin_inventario,
        "reabasto_enterprise_board": reabasto_enterprise_board,
    }


def _reabasto_master_blockers(plan: PlanProduccion | None) -> dict[str, Any]:
    if not plan:
        return {
            "cards": [],
            "detail_rows": [],
            "focus": None,
            "focus_rows": [],
        }

    referencia = f"PLAN_PRODUCCION:{plan.id}"
    solicitudes = list(
        SolicitudCompra.objects.select_related("insumo", "insumo__unidad_base")
        .filter(area=referencia, insumo__isnull=False)
        .order_by("insumo__nombre", "id")
    )
    ordenes = list(
        OrdenCompra.objects.select_related("solicitud__insumo", "solicitud__insumo__unidad_base")
        .filter(referencia=referencia, solicitud__insumo__isnull=False)
        .order_by("solicitud__insumo__nombre", "id")
    )
    recepciones = list(
        RecepcionCompra.objects.select_related("orden__solicitud__insumo", "orden__solicitud__insumo__unidad_base")
        .filter(orden__referencia=referencia, orden__solicitud__insumo__isnull=False)
        .order_by("orden__solicitud__insumo__nombre", "id")
    )

    insumos_by_id: dict[int, Insumo] = {}
    for solicitud in solicitudes:
        if solicitud.insumo_id:
            insumos_by_id[solicitud.insumo_id] = solicitud.insumo
    for orden in ordenes:
        insumo = getattr(getattr(orden, "solicitud", None), "insumo", None)
        if insumo and insumo.id:
            insumos_by_id[insumo.id] = insumo
    for recepcion in recepciones:
        insumo = getattr(getattr(getattr(recepcion, "orden", None), "solicitud", None), "insumo", None)
        if insumo and insumo.id:
            insumos_by_id[insumo.id] = insumo

    groups: dict[str, dict[str, Any]] = {}
    detail_rows: list[dict[str, Any]] = []
    for insumo in sorted(insumos_by_id.values(), key=lambda item: (item.nombre or "").lower()):
        profile = _insumo_erp_readiness(insumo)
        if profile["ready"]:
            continue
        article_class = _insumo_article_class(insumo)
        class_key = article_class["key"]
        class_label = article_class["label"]
        group = groups.setdefault(
            class_key,
            {
                "class_key": class_key,
                "class_label": class_label,
                "count": 0,
                "missing_totals": defaultdict(int),
            },
        )
        group["count"] += 1
        missing_fields = list(profile["missing"])
        for missing_label in missing_fields:
            group["missing_totals"][missing_label] += 1

        primary_missing = missing_fields[0] if missing_fields else None
        action_meta = _enterprise_blocker_action_meta_for_recipes(
            insumo.nombre,
            class_key,
            primary_missing,
            insumo_id=insumo.id,
            usage_scope="recipes",
        )
        detail_rows.append(
            {
                "class_key": class_key,
                "class_label": class_label,
                "insumo_nombre": insumo.nombre,
                "name": insumo.nombre,
                "missing_field": _missing_field_to_filter_key(primary_missing or "") or "maestro",
                "missing": ", ".join(missing_fields),
                "detail": "Completa el maestro para habilitar compras, recepción y abastecimiento CEDIS del plan.",
                "action_label": action_meta["label"],
                "action_detail": action_meta["detail"],
                "action_url": action_meta["url"],
                "edit_url": action_meta["edit_url"],
                "tone": "warning",
            }
        )

    cards: list[dict[str, Any]] = []
    for group in sorted(groups.values(), key=lambda item: (-item["count"], item["class_label"])):
        dominant_label = ""
        dominant_count = 0
        for missing_label, count in dict(group["missing_totals"]).items():
            if count > dominant_count:
                dominant_label = missing_label
                dominant_count = count
        action_label, action_detail = _enterprise_blocker_label_detail_for_missing_recipes(dominant_label)
        query = {
            "tipo_item": group["class_key"],
            "enterprise_status": "incompletos",
            "usage_scope": "recipes",
        }
        filter_key = _missing_field_to_filter_key(dominant_label)
        if filter_key:
            query["missing_field"] = filter_key
        cards.append(
            {
                "key": group["class_key"],
                "class_key": group["class_key"],
                "class_label": group["class_label"],
                "count": group["count"],
                "dominant_label": dominant_label or "maestro incompleto",
                "dominant_count": dominant_count,
                "action_label": action_label,
                "action_detail": action_detail,
                "action_url": reverse("maestros:insumo_list") + f"?{urlencode(query)}",
            }
        )

    focus_rows = list(detail_rows[:3])
    if focus_rows:
        first_focus = focus_rows[0]
        focus = {
            **first_focus,
            "label": f"{first_focus['class_label']} · {first_focus['missing_field']}",
            "summary": (
                f"El flujo CEDIS sigue condicionado por {first_focus['insumo_nombre']} "
                f"({first_focus['missing_field']})."
            ),
            "tone": "warning",
        }
    else:
        focus = {
            "class_label": "Maestro",
            "missing_field": "sin bloqueos",
            "label": "Maestro ERP al día",
            "summary": "No hay artículos del plan bloqueando el flujo CEDIS por faltantes de maestro.",
            "action_label": "Abrir maestro",
            "action_detail": "Puedes revisar el catálogo general para seguimiento preventivo.",
            "action_url": reverse("maestros:insumo_list"),
            "tone": "success",
            "count": 0,
        }

    return {
        "cards": cards[:6],
        "detail_rows": detail_rows[:12],
        "focus": focus,
        "focus_rows": focus_rows,
    }


def _reabasto_daily_control(
    *,
    fecha_operacion: date,
    resumen_cierre: dict[str, Any],
    reabasto_enterprise_board: dict[str, Any],
    consolidado_rows: list[dict[str, Any]],
    stage_key: str = "auto",
    closure_key: str = "auto",
    handoff_key: str = "auto",
    master_focus_key: str = "auto",
) -> dict[str, Any]:
    plan = _find_reabasto_plan(fecha_operacion)
    plan_items = int(plan.items.count()) if plan else 0
    demand_history_summary = _reabasto_demand_history_summary(fecha_operacion)
    master_demand_gate = _plan_master_demand_gate(plan) if plan else None
    master_demand_rows = list((master_demand_gate or {}).get("rows") or [])[:3]
    doc_control = (
        _plan_document_control(
            plan,
            stage_key=stage_key,
            closure_key=closure_key,
            handoff_key=handoff_key,
        )
        if plan
        else None
    )
    generation_gate = _reabasto_generation_gate(
        fecha_operacion=fecha_operacion,
        resumen_cierre=resumen_cierre,
        reabasto_enterprise_board=reabasto_enterprise_board,
        consolidado_rows=consolidado_rows,
        plan=plan,
        doc_control=doc_control,
        demand_history_summary=demand_history_summary,
        master_demand_gate=master_demand_gate,
    )

    stage_label = "Esperando cierres"
    stage_tone = "danger"
    stage_detail = "Todavía faltan cierres o existen sucursales tardías para arrancar producción CEDIS."
    next_action_label = "Revisar cierres"
    next_action_url = "#cierre-sucursales"

    if resumen_cierre["listo_8am"]:
        if (master_demand_gate or {}).get("blockers"):
            stage_label = "Demanda crítica bloqueada"
            stage_tone = "danger"
            stage_detail = str((master_demand_gate or {}).get("detail") or "El reabasto no debe liberarse mientras el maestro crítico siga abierto.")
            next_action_label = str((master_demand_gate or {}).get("action_label") or "Cerrar prioridad crítica")
            next_action_url = str((master_demand_gate or {}).get("action_url") or reverse("maestros:insumo_list"))
        elif reabasto_enterprise_board["blocked_total"] > 0:
            stage_label = "Bloqueos operativos"
            stage_tone = "warning"
            stage_detail = "Hay productos finales bloqueados por receta, empaque, inventario o maestro."
            next_action_label = "Revisar bloqueos"
            next_action_url = "#bloqueos-abastecimiento"
        elif not plan:
            stage_label = "Listo para plan"
            stage_tone = "primary"
            stage_detail = "Los cierres están completos y no hay bloqueos críticos; ya puede generarse el plan CEDIS."
            next_action_label = "Generar plan"
            next_action_url = "#filtro-operativo"
        elif doc_control:
            stage_label = doc_control["stage_label"]
            stage_tone = doc_control["stage_tone"]
            stage_detail = doc_control["stage_detail"]
            next_action_label = doc_control["next_action_label"]
            next_action_url = doc_control["next_action_url"]

    control_cards = [
        {
            "label": "Sucursales pendientes",
            "count": int(resumen_cierre["pendientes"] + resumen_cierre["tardias"]),
            "detail": f"Por validar {resumen_cierre['pendientes']} · Tardías {resumen_cierre['tardias']}",
            "tone": "danger" if (resumen_cierre["pendientes"] + resumen_cierre["tardias"]) else "success",
            "action_label": "Abrir cierres",
            "action_url": "#cierre-sucursales",
        },
        {
            "label": "Productos bloqueados",
            "count": int(reabasto_enterprise_board["blocked_total"]),
            "detail": f"Listos {reabasto_enterprise_board['ready_total']}",
            "tone": "warning" if reabasto_enterprise_board["blocked_total"] else "success",
            "action_label": "Abrir bloqueos",
            "action_url": "#bloqueos-abastecimiento",
        },
        {
            "label": "Plan CEDIS",
            "count": plan_items,
            "detail": plan.nombre if plan else "Sin plan generado",
            "tone": "primary" if plan else "warning",
            "action_label": "Abrir plan" if plan else "Generar plan",
            "action_url": f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}" if plan else "#filtro-operativo",
        },
        {
            "label": "Faltante a producir",
            "count": _quantize_qty(generation_gate["shortage_total"]),
            "detail": f"Solicitado { _quantize_qty(generation_gate['requested_total']) }",
            "tone": "primary" if generation_gate["shortage_total"] > 0 else "success",
            "action_label": "Ver consolidado",
            "action_url": "#consolidado-cedis",
        },
    ]
    if (master_demand_gate or {}).get("blockers"):
        control_cards.insert(
            2,
            {
                "label": "Demanda crítica bloqueada",
                "count": int((master_demand_gate or {}).get("blockers") or 0),
                "detail": str((master_demand_gate or {}).get("detail") or "El maestro crítico sigue abierto."),
                "tone": "danger",
                "action_label": str((master_demand_gate or {}).get("action_label") or "Cerrar prioridad crítica"),
                "action_url": str((master_demand_gate or {}).get("action_url") or reverse("maestros:insumo_list")),
            },
        )

    document_cards: list[dict[str, Any]] = []
    document_stage_rows: list[dict[str, Any]] = []
    document_blocker_rows: list[dict[str, Any]] = []
    pipeline_steps: list[dict[str, Any]] = []
    purchase_gate: dict[str, Any] | None = None
    stage_focus: dict[str, Any] | None = None
    closure_checks: list[dict[str, Any]] = []
    closure_focus: dict[str, Any] | None = None
    closure_focus_rows: list[dict[str, Any]] = []
    handoff_checks: list[dict[str, Any]] = []
    handoff_focus: dict[str, Any] | None = None
    handoff_focus_rows: list[dict[str, Any]] = []
    master_blocker_class_cards: list[dict[str, Any]] = []
    master_blocker_detail_rows: list[dict[str, Any]] = []
    master_focus: dict[str, Any] | None = None
    master_focus_rows: list[dict[str, Any]] = []
    master_summary: dict[str, Any] = {
        "label": "Maestro ERP al día",
        "tone": "success",
        "ready_count": 0,
        "blocked_count": 0,
        "progress_pct": 100,
        "detail": "No hay artículos bloqueando el reabasto CEDIS por faltantes del maestro.",
    }
    trunk_handoff_rows: list[dict[str, Any]] = []
    selected_master_focus_key = "auto"
    reabasto_master_board = _reabasto_master_blockers(plan)
    if doc_control:
        document_cards = doc_control["document_cards"]
        document_stage_rows = doc_control["document_stage_rows"]
        document_blocker_rows = doc_control["document_blocker_rows"]
        pipeline_steps = doc_control["pipeline_steps"]
        purchase_gate = doc_control["purchase_gate"]
        stage_focus = doc_control["stage_focus"]
        closure_checks = doc_control["closure_checks"]
        closure_focus = doc_control["closure_focus"]
        closure_focus_rows = doc_control.get("closure_focus_rows") or []
        handoff_checks = doc_control["handoff_checks"]
        handoff_focus = doc_control["handoff_focus"]
        handoff_focus_rows = doc_control.get("handoff_focus_rows") or []
        master_cards = list(reabasto_master_board.get("cards") or [])
        valid_master_focus_keys = {str(card.get("key") or "").strip() for card in master_cards if card.get("key")}
        selected_master_focus_key = master_focus_key if master_focus_key in valid_master_focus_keys else "auto"
        focus_base = reverse("recetas:reabasto_cedis") + f"?{urlencode({'fecha': fecha_operacion.isoformat()})}"
        for card in master_cards:
            card_key = str(card.get("key") or "").strip()
            card["focus_url"] = f"{focus_base}&master_focus_key={urlencode({'k': card_key})[2:]}"
            card["is_active"] = selected_master_focus_key != "auto" and card_key == selected_master_focus_key
        master_detail_source = list(reabasto_master_board.get("detail_rows") or [])
        master_blocker_detail_rows = (
            [row for row in master_detail_source if str(row.get("class_key") or "").strip() == selected_master_focus_key]
            if selected_master_focus_key != "auto"
            else master_detail_source
        )
        master_blocker_class_cards = master_cards
        master_focus_rows = list(master_blocker_detail_rows[:3])
        master_focus = reabasto_master_board.get("focus")
        if master_focus and selected_master_focus_key != "auto" and master_focus_rows:
            first_master_focus = master_focus_rows[0]
            master_focus = {
                **master_focus,
                **first_master_focus,
                "label": f"{first_master_focus['class_label']} · {first_master_focus['missing_field']}",
                "summary": (
                    f"El flujo CEDIS sigue condicionado por {first_master_focus['insumo_nombre']} "
                    f"({first_master_focus['missing_field']})."
                ),
            }
        master_blocked_total = sum(int(card.get("count") or 0) for card in master_cards) or len(master_detail_source)
        master_ready_total = 0 if master_blocked_total else int(reabasto_master_board.get("ready_for_purchase_total") or 0)
        master_total = master_ready_total + master_blocked_total
        master_progress_pct = int(round((master_ready_total / master_total) * 100)) if master_total else 100
        if master_blocked_total:
            dominant_card = master_cards[0] if master_cards else None
            dominant_text = ""
            if dominant_card:
                dominant_text = (
                    f" Bloqueo dominante: {dominant_card.get('class_label', 'Artículo')} · "
                    f"{dominant_card.get('dominant_label', 'maestro incompleto')}."
                )
            master_summary = {
                "label": "Maestro ERP con bloqueos",
                "tone": "warning",
                "ready_count": master_ready_total,
                "blocked_count": master_blocked_total,
                "progress_pct": master_progress_pct,
                "detail": (
                    f"{master_blocked_total} artículo(s) siguen condicionando el reabasto por datos incompletos."
                    f"{dominant_text}"
                ),
            }
        else:
            master_summary = {
                "label": "Maestro ERP al día",
                "tone": "success",
                "ready_count": master_ready_total,
                "blocked_count": 0,
                "progress_pct": master_progress_pct,
                "detail": "El maestro no está bloqueando el reabasto CEDIS para esta fecha.",
            }

    generation_blockers = sum(1 for item in generation_gate.get("checks", []) if not item.get("is_ready"))
    critical_master_open = bool((master_demand_gate or {}).get("blockers") or 0)
    plan_ready = bool(plan and generation_gate.get("can_generate_plan"))
    compras_ready = bool(plan and generation_gate.get("can_generate_compras"))
    inventario_ready = bool(
        plan
        and int(reabasto_enterprise_board["blocked_total"]) == 0
        and int(master_summary.get("blocked_count") or 0) == 0
    )
    trunk_handoff_rows = [
        {
            "label": "Sucursales / Plan",
            "owner": "Ventas / Planeación / Producción",
            "status": "Crítico" if critical_master_open else "Listo para operar" if plan_ready else "Bloqueado",
            "tone": "danger" if critical_master_open else "success" if plan_ready else "warning",
            "blockers": generation_blockers,
            "completion": 12 if critical_master_open else 100 if plan_ready else max(0, 100 - (generation_blockers * 12)),
            "depends_on": "Cierres completos + faltante real + plan generado",
            "exit_criteria": "CEDIS debe contar con cierres completos y un plan de producción válido para arrancar.",
            "detail": (
                "CEDIS no debe liberar el día mientras siga abierta la demanda crítica bloqueada por maestro."
                if critical_master_open
                else "CEDIS ya tiene cierres completos y plan operativo para arrancar."
                if plan_ready
                else "Todavía faltan cierres, faltante validado o plan generado para arrancar el día."
            ),
            "next_step": str((master_demand_gate or {}).get("next_step") or next_action_label) if critical_master_open else next_action_label,
            "url": str((master_demand_gate or {}).get("action_url") or next_action_url) if critical_master_open else next_action_url,
            "cta": str((master_demand_gate or {}).get("action_label") or next_action_label) if critical_master_open else next_action_label,
        },
        {
            "label": "Compras documentales",
            "owner": "Compras / Almacén",
            "status": "Crítico" if critical_master_open else "Listo para operar" if compras_ready else "Bloqueado",
            "tone": "danger" if critical_master_open else "success" if compras_ready else "warning",
            "blockers": int((purchase_gate or {}).get("blocked") or 0) if purchase_gate else 1,
            "completion": 12 if critical_master_open else 100 if compras_ready else int(master_summary.get("progress_pct") or 0),
            "depends_on": "Solicitudes + órdenes + recepciones sin bloqueo",
            "exit_criteria": "El abastecimiento documental debe quedar emitido y sin bloqueos para surtir a tiempo.",
            "detail": (
                "Compras no debe avanzar mientras siga abierta la demanda crítica bloqueada por maestro."
                if critical_master_open
                else "El flujo documental ya puede operar para abastecer el reabasto del día."
                if compras_ready
                else (purchase_gate or {}).get("detail") or "Compras sigue condicionado por bloqueos documentales o de maestro."
            ),
            "next_step": str((master_demand_gate or {}).get("next_step") or (purchase_gate or {}).get("cta_label") or "Abrir plan") if critical_master_open else (purchase_gate or {}).get("cta_label") or "Abrir plan",
            "url": str((master_demand_gate or {}).get("action_url") or (purchase_gate or {}).get("cta_url") or (f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}" if plan else "#filtro-operativo")) if critical_master_open else (purchase_gate or {}).get("cta_url") or (f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}" if plan else "#filtro-operativo"),
            "cta": str((master_demand_gate or {}).get("action_label") or (purchase_gate or {}).get("cta_label") or "Abrir plan") if critical_master_open else (purchase_gate or {}).get("cta_label") or "Abrir plan",
        },
        {
            "label": "Inventario / Reabasto",
            "owner": "Inventario / CEDIS",
            "status": "Listo para operar" if inventario_ready else "Bloqueado",
            "tone": "success" if inventario_ready else "warning",
            "blockers": int(reabasto_enterprise_board["blocked_total"]) + int(master_summary.get("blocked_count") or 0),
            "completion": 100 if inventario_ready else max(0, 100 - ((int(reabasto_enterprise_board["blocked_total"]) + int(master_summary.get("blocked_count") or 0) or 1) * 10)),
            "depends_on": "Productos desbloqueados + maestro al día + inventario CEDIS cargado",
            "exit_criteria": "CEDIS debe poder producir y surtir sin bloqueos de receta, empaque o maestro.",
            "detail": (
                "Inventario y reabasto ya pueden sostener la operación diaria de CEDIS."
                if inventario_ready
                else "CEDIS todavía tiene productos bloqueados o artículos incompletos en maestro."
            ),
            "next_step": "Ver bloqueos" if not inventario_ready else "Abrir consolidado",
            "url": "#bloqueos-abastecimiento" if not inventario_ready else "#consolidado-cedis",
            "cta": "Ver bloqueos" if not inventario_ready else "Abrir consolidado",
        },
    ]

    return {
        "stage_label": stage_label,
        "stage_tone": stage_tone,
        "stage_detail": stage_detail,
        "next_action_label": next_action_label,
        "next_action_url": next_action_url,
        "plan": plan,
        "control_cards": control_cards,
        "document_cards": document_cards,
        "document_stage_rows": document_stage_rows,
        "document_blocker_rows": document_blocker_rows,
        "pipeline_steps": pipeline_steps,
        "purchase_gate": purchase_gate,
        "stage_focus": stage_focus,
        "closure_checks": closure_checks,
        "closure_focus": closure_focus,
        "closure_focus_rows": closure_focus_rows,
        "handoff_checks": handoff_checks,
        "handoff_focus": handoff_focus,
        "handoff_focus_rows": handoff_focus_rows,
        "master_blocker_class_cards": master_blocker_class_cards,
        "master_blocker_detail_rows": master_blocker_detail_rows,
        "master_focus": master_focus,
        "master_focus_rows": master_focus_rows,
        "master_summary": master_summary,
        "master_demand_gate": master_demand_gate,
        "master_demand_rows": master_demand_rows,
        "selected_master_focus_key": selected_master_focus_key,
        "generation_gate": generation_gate,
        "trunk_handoff_rows": trunk_handoff_rows,
        "trunk_handoff_summary": _trunk_handoff_summary(
            trunk_handoff_rows,
            owner="CEDIS / Compras / Inventario",
            fallback_url=reverse("recetas:reabasto_cedis") + f"?{urlencode({'fecha': fecha_operacion.isoformat()})}",
        ),
    }


def _reabasto_daily_decisions(
    *,
    demand_history_summary: dict[str, Any] | None,
    daily_control: dict[str, Any],
) -> list[dict[str, object]]:
    decisions: list[dict[str, object]] = []
    generation_gate = dict(daily_control.get("generation_gate") or {})
    demand_gate = dict(generation_gate.get("demand_gate") or {})
    master_demand_gate = dict(daily_control.get("master_demand_gate") or {})
    purchase_gate = dict(daily_control.get("purchase_gate") or {})

    def push(priority: int, tone: str, title: str, detail: str, url: str, cta: str) -> None:
        decisions.append(
            {
                "priority": priority,
                "tone": tone,
                "title": title,
                "detail": detail,
                "url": url,
                "cta": cta,
            }
        )

    if master_demand_gate and int(master_demand_gate.get("blockers") or 0) > 0:
        push(
            100,
            str(master_demand_gate.get("tone") or "danger"),
            "Cerrar maestro crítico del reabasto",
            str(master_demand_gate.get("detail") or "Hay artículos críticos bloqueando el plan CEDIS."),
            str(master_demand_gate.get("action_url") or reverse("maestros:insumo_list")),
            str(master_demand_gate.get("action_label") or "Abrir maestro"),
        )

    if demand_gate and str(demand_gate.get("tone") or "") == "danger":
        push(
            95,
            "danger",
            "Validar base comparable de demanda",
            str(demand_gate.get("detail") or "La base comparable no es suficiente para generar el reabasto con confianza."),
            str(demand_gate.get("action_url") or "#historico-demanda"),
            str(demand_gate.get("action_label") or "Revisar demanda"),
        )

    if demand_history_summary and int(demand_history_summary.get("sample_days") or 0) == 0:
        push(
            90,
            "danger",
            "Construir referencia histórica",
            "No hay días comparables previos para soportar el reabasto. Conviene revisar captura y demanda antes de generar plan.",
            "#historico-demanda",
            "Ver histórico",
        )

    if str(daily_control.get("stage_tone") or "") == "danger":
        push(
            85,
            str(daily_control.get("stage_tone") or "danger"),
            "Completar cierres de sucursal",
            str(daily_control.get("stage_detail") or "El arranque CEDIS sigue esperando cierres de sucursal."),
            str(daily_control.get("next_action_url") or reverse("recetas:reabasto_cedis_captura")),
            str(daily_control.get("next_action_label") or "Abrir cierres"),
        )

    if purchase_gate and str(purchase_gate.get("tone") or "") != "success":
        push(
            80,
            str(purchase_gate.get("tone") or "warning"),
            "Destrabar flujo de compras",
            str(purchase_gate.get("detail") or "El flujo documental de compras sigue abierto o bloqueado para este reabasto."),
            str(purchase_gate.get("cta_url") or reverse("compras:solicitudes")),
            str(purchase_gate.get("cta_label") or "Abrir compras"),
        )

    if not decisions:
        push(
            10,
            "success",
            "Reabasto sin alertas dominantes",
            "Cierres, demanda, maestro y compras están en condición operable para el arranque CEDIS.",
            reverse("recetas:reabasto_cedis"),
            "Actualizar reabasto",
        )

    decisions.sort(key=lambda item: int(item.get("priority") or 0), reverse=True)
    return decisions[:5]


def _reabasto_branch_priority_rows(
    *,
    fecha_operacion: date,
    sucursales: list[Sucursal],
    resumen_cierre: dict[str, Any],
    demand_history_summary: dict[str, Any] | None,
) -> list[dict[str, object]]:
    solicitud_map = {
        item.sucursal_id: item
        for item in SolicitudReabastoCedis.objects.filter(fecha_operacion=fecha_operacion, sucursal__in=sucursales)
        .select_related("sucursal")
    }
    demand_branch_map = {
        row.get("sucursal__codigo"): row
        for row in list((demand_history_summary or {}).get("top_branches") or [])
    }
    cierre_map = {row["sucursal"].id: row for row in list(resumen_cierre.get("detalle") or [])}

    rows: list[dict[str, object]] = []
    for sucursal in sucursales:
        cierre = cierre_map.get(sucursal.id) or {}
        solicitud = solicitud_map.get(sucursal.id)
        demand_row = demand_branch_map.get(sucursal.codigo) or {}
        lineas = list(solicitud.lineas.all()) if solicitud else []
        solicitado_total = sum((_to_decimal_safe(linea.solicitado) for linea in lineas), Decimal("0"))
        sugerido_total = sum((_to_decimal_safe(linea.sugerido) for linea in lineas), Decimal("0"))
        dominant_line = None
        dominant_units = Decimal("0")
        for linea in lineas:
            candidate_units = max(_to_decimal_safe(linea.solicitado), _to_decimal_safe(linea.sugerido))
            if candidate_units > dominant_units:
                dominant_line = linea
                dominant_units = candidate_units
        pendiente_envio = bool(cierre.get("semaforo") in {"rojo", "amarillo"})
        demand_units = Decimal(str(demand_row.get("total") or 0))
        if pendiente_envio and demand_units > 0:
            tone = "danger"
            status = "Prioridad crítica"
            detail = f"Cierre pendiente/tardío con {demand_units:.0f} unidades históricas comparables."
            priority_score = demand_units + Decimal("200")
        elif pendiente_envio:
            tone = "warning"
            status = "Cierre pendiente"
            detail = "La sucursal todavía no deja un cierre confiable para el arranque CEDIS."
            priority_score = Decimal("120")
        elif sugerido_total > 0 or solicitado_total > 0:
            tone = "primary"
            status = "Lista para surtir"
            detail = f"Sugerido { _quantize_qty(sugerido_total) } · Solicitado { _quantize_qty(solicitado_total) }."
            priority_score = max(sugerido_total, solicitado_total)
        else:
            tone = "success"
            status = "Sin presión visible"
            detail = "No trae cierre atrasado ni faltante sugerido relevante para hoy."
            priority_score = Decimal("0")

        rows.append(
            {
                "sucursal_codigo": sucursal.codigo,
                "sucursal_nombre": sucursal.nombre,
                "status": status,
                "tone": tone,
                "detail": detail,
                "historico_units": demand_units,
                "sugerido_total": _quantize_qty(sugerido_total),
                "solicitado_total": _quantize_qty(solicitado_total),
                "dominant_recipe_id": int(getattr(dominant_line, "receta_id", 0) or 0),
                "dominant_recipe_name": (
                    getattr(getattr(dominant_line, "receta", None), "nombre", "") or "Producto"
                ),
                "dominant_recipe_units": _quantize_qty(dominant_units),
                "action_url": _reabasto_redirect(fecha_operacion, sucursal.id, capture_only=False),
                "action_label": "Abrir sucursal",
                "priority_score": priority_score,
            }
        )

    tone_order = {"danger": 0, "warning": 1, "primary": 2, "success": 3}
    rows.sort(
        key=lambda item: (
            tone_order.get(str(item.get("tone") or ""), 9),
            -float(item.get("priority_score") or 0),
            str(item.get("sucursal_codigo") or ""),
        )
    )
    return rows[:6]


def _reabasto_branch_supply_rows(
    *,
    fecha_operacion: date,
    branch_priority_rows: list[dict[str, object]],
    limit: int = 6,
) -> list[dict[str, object]]:
    if not branch_priority_rows:
        return []

    recipe_ids = {
        int(row.get("dominant_recipe_id") or 0)
        for row in branch_priority_rows
        if int(row.get("dominant_recipe_id") or 0) > 0
    }
    if not recipe_ids:
        return []

    lineas = list(
        LineaReceta.objects.filter(receta_id__in=recipe_ids, insumo_id__isnull=False)
        .exclude(tipo_linea=LineaReceta.TIPO_SUBSECCION)
        .select_related("insumo", "receta", "insumo__unidad_base")
    )
    if not lineas:
        return []

    canonical_by_line: dict[int, Insumo] = {}
    canonical_ids: set[int] = set()
    lineas_by_recipe: dict[int, list[LineaReceta]] = defaultdict(list)
    for linea in lineas:
        if not linea.insumo:
            continue
        canonical = canonical_insumo(linea.insumo) or linea.insumo
        canonical_by_line[linea.id] = canonical
        canonical_ids.add(canonical.id)
        lineas_by_recipe[int(linea.receta_id)].append(linea)

    existencia_map = {
        int(item.insumo_id): item
        for item in ExistenciaInsumo.objects.filter(insumo_id__in=canonical_ids).select_related("insumo")
    }

    rows: list[dict[str, object]] = []
    for branch_row in branch_priority_rows:
        recipe_id = int(branch_row.get("dominant_recipe_id") or 0)
        recipe_units = _to_decimal_safe(branch_row.get("dominant_recipe_units"))
        if recipe_id <= 0 or recipe_units <= 0:
            continue

        best_candidate: dict[str, object] | None = None
        best_score = Decimal("-1")
        for linea in lineas_by_recipe.get(recipe_id, []):
            canonical = canonical_by_line.get(linea.id)
            if canonical is None:
                continue

            required_qty = _to_decimal_safe(linea.cantidad) * recipe_units
            if required_qty <= 0:
                continue

            existencia = existencia_map.get(canonical.id)
            stock_actual = _to_decimal_safe(getattr(existencia, "stock_actual", 0))
            shortage = max(required_qty - stock_actual, Decimal("0"))
            readiness = _insumo_erp_readiness(canonical)
            missing = list(readiness.get("missing") or [])
            latest_cost = _latest_cost_for_insumo(canonical)
            missing_cost = latest_cost is None
            score = (shortage * Decimal("100")) + (Decimal(str(len(missing))) * Decimal("50")) + required_qty
            if missing_cost:
                score += Decimal("25")
            if score > best_score:
                best_score = score
                best_candidate = {
                    "insumo_nombre": canonical.nombre,
                    "required_qty": _quantize_qty(required_qty),
                    "stock_actual": _quantize_qty(stock_actual),
                    "shortage": _quantize_qty(shortage),
                    "master_missing": missing,
                    "missing_cost": missing_cost,
                    "unidad": canonical.unidad_base.codigo if getattr(canonical, "unidad_base_id", None) else "-",
                    "action_url": reverse("maestros:insumo_update", args=[canonical.id]),
                    "action_label": "Asegurar insumo",
                }

        if not best_candidate:
            continue

        rows.append(
            {
                "sucursal_codigo": branch_row.get("sucursal_codigo") or "",
                "sucursal_nombre": branch_row.get("sucursal_nombre") or "Sucursal",
                "dominant_recipe_name": branch_row.get("dominant_recipe_name") or "Producto",
                "insumo_nombre": best_candidate["insumo_nombre"],
                "required_qty": best_candidate["required_qty"],
                "stock_actual": best_candidate["stock_actual"],
                "shortage": best_candidate["shortage"],
                "master_missing": best_candidate["master_missing"],
                "missing_cost": best_candidate["missing_cost"],
                "unidad": best_candidate["unidad"],
                "action_url": best_candidate["action_url"],
                "action_label": best_candidate["action_label"],
                "priority_score": best_score,
            }
        )

    rows.sort(
        key=lambda item: (
            Decimal(str(item.get("shortage") or 0)),
            Decimal(str(len(item.get("master_missing") or []))),
            Decimal(str(item.get("required_qty") or 0)),
        ),
        reverse=True,
    )
    return rows[:limit]


def _reabasto_generation_gate(
    *,
    fecha_operacion: date,
    resumen_cierre: dict[str, Any],
    reabasto_enterprise_board: dict[str, Any],
    consolidado_rows: list[dict[str, Any]],
    plan: PlanProduccion | None,
    doc_control: dict[str, Any] | None,
    demand_history_summary: dict[str, Any] | None = None,
    master_demand_gate: dict[str, Any] | None = None,
) -> dict[str, Any]:
    requested_total = sum((_to_decimal_safe(row.get("total_solicitado")) for row in consolidado_rows), Decimal("0"))
    shortage_total = sum((_to_decimal_safe(row.get("cedis_faltante_producir")) for row in consolidado_rows), Decimal("0"))
    doc_blocked_total = int((doc_control or {}).get("blocked_total") or 0)
    doc_total = 0
    if doc_control:
        doc_total = int(doc_control.get("solicitudes_total") or 0) + int(doc_control.get("ordenes_total") or 0) + int(doc_control.get("recepciones_total") or 0)
    demand_gate = _commercial_signal_gate(
        demand_history_summary,
        context_label="el reabasto diario",
        action_url="#historico-demanda",
        action_label="Revisar base histórica",
    )
    master_demand_ready = not bool((master_demand_gate or {}).get("blockers") or 0)

    checks = [
        {
            "label": "Cierres de sucursal completos",
            "is_ready": bool(resumen_cierre["listo_8am"]),
            "detail": (
                "Todas las sucursales enviaron en tiempo para arranque 8:00 AM."
                if resumen_cierre["listo_8am"]
                else f"Por atender {resumen_cierre['pendientes']} · Tardías {resumen_cierre['tardias']}"
            ),
            "action_label": "Abrir cierres",
            "action_url": "#cierre-sucursales",
        },
        {
            "label": "Productos finales sin bloqueo",
            "is_ready": int(reabasto_enterprise_board["blocked_total"]) == 0,
            "detail": (
                "No hay bloqueos por receta, empaque, inventario ni maestro."
                if int(reabasto_enterprise_board["blocked_total"]) == 0
                else f"Bloqueadas {reabasto_enterprise_board['blocked_total']} · Listas {reabasto_enterprise_board['ready_total']}"
            ),
            "action_label": "Ver bloqueos",
            "action_url": "#bloqueos-abastecimiento",
        },
        {
            "label": "Base histórica comparable",
            "is_ready": bool(demand_gate["is_ready"]),
            "detail": str(demand_gate["detail"]),
            "action_label": str(demand_gate["action_label"]),
            "action_url": str(demand_gate["action_url"]),
        },
        {
            "label": "Maestro crítico del plan cerrado",
            "is_ready": master_demand_ready,
            "detail": (
                "No hay artículos críticos por demanda bloqueando el plan CEDIS."
                if master_demand_ready
                else str((master_demand_gate or {}).get("detail") or "El plan sigue bloqueado por artículos críticos del maestro.")
            ),
            "action_label": str((master_demand_gate or {}).get("action_label") or "Cerrar prioridad crítica"),
            "action_url": str((master_demand_gate or {}).get("action_url") or reverse("maestros:insumo_list")),
        },
        {
            "label": "Faltante a producir detectado",
            "is_ready": shortage_total > 0,
            "detail": (
                f"Faltante CEDIS { _quantize_qty(shortage_total) }"
                if shortage_total > 0
                else "No hay faltante a producir; no corresponde generar plan."
            ),
            "action_label": "Ver consolidado",
            "action_url": "#consolidado-cedis",
        },
        {
            "label": "Plan CEDIS generado",
            "is_ready": bool(plan),
            "detail": (
                f"{plan.nombre} · {int(plan.items.count())} renglones"
                if plan
                else "Todavía no existe plan para esta fecha."
            ),
            "action_label": "Abrir plan" if plan else "Generar plan",
            "action_url": f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}" if plan else "#filtro-operativo",
        },
        {
            "label": "Flujo de compras sin bloqueos",
            "is_ready": bool(plan) and doc_total > 0 and doc_blocked_total == 0,
            "detail": (
                "Sin documentos bloqueados en solicitudes, órdenes o recepciones."
                if plan and doc_total > 0 and doc_blocked_total == 0
                else f"Bloqueos documentales {doc_blocked_total}"
                if plan and doc_total > 0
                else (
                    "Todavía no existen documentos de compras para este plan."
                    if plan
                    else "Primero debe existir un plan CEDIS antes de generar compras."
                )
            ),
            "action_label": (
                (doc_control or {}).get("next_action_label")
                if doc_control
                else ("Abrir plan" if plan else "Generar plan")
            ),
            "action_url": (
                (doc_control or {}).get("next_action_url")
                if doc_control
                else (f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}" if plan else "#filtro-operativo")
            ),
        },
    ]

    can_generate_plan = bool(checks[0]["is_ready"] and checks[1]["is_ready"] and checks[2]["is_ready"] and checks[3]["is_ready"] and checks[4]["is_ready"])
    can_generate_compras = bool(can_generate_plan and checks[6]["is_ready"])

    if can_generate_compras:
        overall_label = "Listo para generar"
        overall_tone = "success"
        overall_detail = "Puede generarse o regenerarse el flujo documental de compras desde el faltante consolidado."
    elif can_generate_plan:
        overall_label = "Listo para plan"
        overall_tone = "primary"
        overall_detail = "El plan puede generarse, pero compras sigue condicionado por bloqueos documentales existentes."
    else:
        overall_label = "Bloqueado"
        overall_tone = "danger"
        overall_detail = "Primero corrige cierres, bloqueos operativos o faltante inexistente antes de generar documentos."

    return {
        "requested_total": requested_total,
        "shortage_total": shortage_total,
        "doc_total": doc_total,
        "checks": checks,
        "can_generate_plan": can_generate_plan,
        "can_generate_compras": can_generate_compras,
        "overall_label": overall_label,
        "overall_tone": overall_tone,
        "overall_detail": overall_detail,
        "demand_gate": demand_gate,
    }


def _reabasto_generation_blocker_message(generation_gate: dict[str, Any], *, target: str) -> str:
    failing = [item["label"] for item in generation_gate.get("checks", []) if not item.get("is_ready")]
    if target == "compras" and generation_gate.get("can_generate_plan") and not generation_gate.get("can_generate_compras"):
        return "No se puede generar compras todavía: el plan está listo, pero el flujo documental actual tiene bloqueos abiertos."
    if failing:
        return f"No se puede generar {target} todavía: {', '.join(failing)}."
    return f"No se puede generar {target} todavía por validaciones enterprise en revisión."


def _log_reabasto_gate_block(user, *, fecha_operacion: date, target: str, generation_gate: dict[str, Any]) -> None:
    failing = [item["label"] for item in generation_gate.get("checks", []) if not item.get("is_ready")]
    log_event(
        user,
        "BLOCKED",
        "recetas.ReabastoCedis",
        None,
        {
            "source": "REABASTO_CEDIS_GATE",
            "target": target,
            "fecha_operacion": fecha_operacion.isoformat(),
            "overall_label": generation_gate.get("overall_label"),
            "overall_tone": generation_gate.get("overall_tone"),
            "requested_total": float(_to_decimal_safe(generation_gate.get("requested_total"))),
            "shortage_total": float(_to_decimal_safe(generation_gate.get("shortage_total"))),
            "doc_total": int(generation_gate.get("doc_total") or 0),
            "failing_checks": failing,
        },
    )


def _recipes_critical_path_rows(
    rows: list[dict[str, object]],
    *,
    owner: str = "Recetas / Operación",
    fallback_url: str,
) -> list[dict[str, object]]:
    severity_order = {"danger": 0, "warning": 1, "success": 2, "primary": 3}
    normalized_rows: list[dict[str, object]] = []
    for row in rows:
        open_count = int(
            row.get("open_count")
            if row.get("open_count") is not None
            else row.get("open")
            if row.get("open") is not None
            else row.get("blocked")
            if row.get("blocked") is not None
            else row.get("count")
            or 0
        )
        completion = int(
            row.get("progress_pct")
            if row.get("progress_pct") is not None
            else row.get("completion")
            if row.get("completion") is not None
            else 0
        )
        tone = str(row.get("semaphore_tone") or row.get("tone") or ("danger" if open_count else "success"))
        normalized_rows.append(
            {
                "title": row.get("label") or row.get("title") or row.get("front") or "Tramo de recetas",
                "owner": row.get("owner") or owner,
                "status": row.get("status") or row.get("semaphore_label") or ("En revisión" if open_count else "Cerrado"),
                "tone": tone,
                "count": open_count,
                "completion": completion,
                "depends_on": row.get("depends_on") or ("Tramo previo del documento" if row.get("step") else "Inicio del flujo"),
                "dependency_status": row.get("dependency_status") or row.get("detail") or "Sin dependencia registrada",
                "detail": row.get("detail", ""),
                "next_step": row.get("next_step") or row.get("cta") or row.get("action_detail") or "Revisar tramo",
                "url": row.get("action_url") or row.get("action_href") or row.get("url") or fallback_url,
                "cta": row.get("action_label") or row.get("cta") or "Abrir",
            }
        )

    ranked = sorted(
        normalized_rows,
        key=lambda item: (
            severity_order.get(str(item.get("tone") or "warning"), 9),
            -int(item.get("count") or 0),
            int(item.get("completion") or 0),
        ),
    )
    critical_rows: list[dict[str, object]] = []
    for index, item in enumerate(ranked[:4], start=1):
        critical_rows.append(
            {
                "rank": f"R{index}",
                "title": item["title"],
                "owner": item["owner"],
                "status": item["status"],
                "tone": item["tone"],
                "count": item["count"],
                "completion": item["completion"],
                "depends_on": item["depends_on"],
                "dependency_status": item["dependency_status"],
                "detail": item["detail"],
                "next_step": item["next_step"],
                "url": item["url"],
                "cta": item["cta"],
            }
        )
    return critical_rows


def _recipes_executive_radar_rows(
    rows: list[dict[str, object]],
    *,
    owner: str = "Recetas / Operación",
    fallback_url: str,
) -> list[dict[str, object]]:
    radar_rows: list[dict[str, object]] = []
    for row in rows[:4]:
        blockers = int(
            row.get("open_count")
            if row.get("open_count") is not None
            else row.get("open")
            if row.get("open") is not None
            else row.get("blocked")
            if row.get("blocked") is not None
            else row.get("count")
            or 0
        )
        completion = int(
            row.get("progress_pct")
            if row.get("progress_pct") is not None
            else row.get("completion")
            if row.get("completion") is not None
            else 0
        )
        tone = str(row.get("semaphore_tone") or row.get("tone") or ("danger" if blockers else "success"))
        if blockers <= 0 and completion >= 90:
            status = "Controlado"
            dominant_blocker = "Sin bloqueo activo"
        elif completion >= 50:
            status = "En seguimiento"
            dominant_blocker = row.get("detail", "") or "Brecha operativa en seguimiento"
        else:
            status = "Con bloqueo"
            dominant_blocker = row.get("detail", "") or "Bloqueo operativo abierto"
        radar_rows.append(
            {
                "phase": row.get("label") or row.get("title") or row.get("front") or "Frente de recetas",
                "owner": row.get("owner") or owner,
                "status": status,
                "tone": tone,
                "blockers": blockers,
                "progress_pct": completion,
                "dominant_blocker": dominant_blocker,
                "depends_on": row.get("depends_on") or "Inicio del flujo",
                "dependency_status": row.get("dependency_status") or row.get("detail") or "Sin dependencia registrada",
                "next_step": row.get("next_step") or row.get("cta") or row.get("action_detail") or "Revisar tramo",
                "url": row.get("action_url") or row.get("action_href") or row.get("url") or fallback_url,
                "cta": row.get("action_label") or row.get("cta") or "Abrir",
            }
        )
    return radar_rows


def _trunk_handoff_summary(
    rows: list[dict[str, object]],
    *,
    owner: str,
    fallback_url: str,
) -> dict[str, object] | None:
    if not rows:
        return None
    normalized: list[dict[str, object]] = []
    for row in rows:
        blockers = int(
            row.get("blockers")
            if row.get("blockers") is not None
            else row.get("open_count")
            if row.get("open_count") is not None
            else row.get("open")
            if row.get("open") is not None
            else row.get("count")
            or 0
        )
        completion = int(
            row.get("completion")
            if row.get("completion") is not None
            else row.get("progress_pct")
            if row.get("progress_pct") is not None
            else 0
        )
        normalized.append(
            {
                "label": row.get("label") or row.get("title") or "Tramo ERP",
                "blockers": blockers,
                "completion": completion,
                "next_step": row.get("next_step") or row.get("cta") or "Revisar tramo",
                "url": row.get("url") or row.get("action_url") or fallback_url,
                "cta": row.get("cta") or row.get("action_label") or "Abrir",
            }
        )
    avg_completion = int(round(sum(int(item["completion"]) for item in normalized) / len(normalized)))
    total_blockers = sum(int(item["blockers"]) for item in normalized)
    priority = sorted(normalized, key=lambda item: (-int(item["blockers"]), int(item["completion"])))[0]
    tone = "success" if total_blockers == 0 and avg_completion >= 100 else "warning" if total_blockers == 0 else "danger"
    status = "Controlado" if tone == "success" else "Seguimiento" if tone == "warning" else "Crítico"
    return {
        "owner": owner,
        "status": status,
        "tone": tone,
        "completion": avg_completion,
        "blockers": total_blockers,
        "next_step": priority["next_step"],
        "url": priority["url"],
        "cta": priority["cta"],
        "priority_label": priority["label"],
        "segments": len(normalized),
    }


def _plan_daily_decisions(
    *,
    plan_actual: PlanProduccion | None,
    demand_gate_summary: dict[str, Any] | None,
    master_demand_gate_summary: dict[str, Any] | None,
    ventas_historicas_summary: dict[str, Any] | None,
    document_control: dict[str, Any] | None,
) -> list[dict[str, object]]:
    decisions: list[dict[str, object]] = []

    def push(priority: int, tone: str, title: str, detail: str, url: str, cta: str) -> None:
        decisions.append(
            {
                "priority": priority,
                "tone": tone,
                "title": title,
                "detail": detail,
                "url": url,
                "cta": cta,
            }
        )

    if master_demand_gate_summary and int(master_demand_gate_summary.get("blockers") or 0) > 0:
        push(
            100,
            str(master_demand_gate_summary.get("tone") or "danger"),
            "Cerrar maestro crítico del plan",
            str(master_demand_gate_summary.get("detail") or "Hay artículos maestros críticos reteniendo la liberación del plan."),
            str(master_demand_gate_summary.get("action_url") or reverse("maestros:insumo_list")),
            str(master_demand_gate_summary.get("action_label") or "Abrir maestro"),
        )

    if demand_gate_summary and str(demand_gate_summary.get("tone") or "") == "danger":
        push(
            95,
            "danger",
            "Validar base comercial del plan",
            str(demand_gate_summary.get("detail") or "La base comercial del plan es frágil."),
            str(demand_gate_summary.get("action_url") or "#plan-pronosticos"),
            str(demand_gate_summary.get("action_label") or "Abrir pronóstico"),
        )

    if ventas_historicas_summary and int(ventas_historicas_summary.get("missing_days") or 0) > 0:
        push(
            88,
            "warning",
            "Completar histórico de ventas",
            (
                f"Faltan {ventas_historicas_summary.get('missing_days', 0)} días dentro del rango histórico. "
                "Eso debilita el soporte estadístico del plan."
            ),
            "#plan-pronosticos",
            "Ver base estadística",
        )

    if not plan_actual:
        push(
            92,
            "danger",
            "Crear plan operativo",
            "No hay un plan seleccionado. Primero define el documento rector antes de explotar demanda y compras.",
            reverse("recetas:plan_produccion"),
            "Crear plan",
        )

    if document_control and int(document_control.get("blocked_total") or 0) > 0:
        push(
            82,
            "warning",
            "Destrabar flujo documental",
            f"Hay {document_control.get('blocked_total', 0)} bloqueos documentales en solicitudes, órdenes o recepciones del plan.",
            str(document_control.get("next_action_url") or reverse("compras:solicitudes")),
            str(document_control.get("next_action_label") or "Abrir compras"),
        )

    if document_control and int(document_control.get("recepciones_abiertas_total") or 0) > 0:
        push(
            75,
            "warning",
            "Cerrar recepciones abiertas",
            f"Hay {document_control.get('recepciones_abiertas_total', 0)} recepciones abiertas que siguen impactando el cierre operativo del plan.",
            reverse("compras:recepciones"),
            "Abrir recepciones",
        )

    if not decisions:
        push(
            10,
            "success",
            "Plan sin alertas dominantes",
            "El plan tiene base comercial, maestro y flujo documental en condición operable.",
            reverse("recetas:plan_produccion"),
            "Actualizar plan",
        )

    decisions.sort(key=lambda item: int(item.get("priority") or 0), reverse=True)
    return decisions[:5]


def _plan_branch_priority_rows(
    *,
    plan_actual: PlanProduccion | None,
    periodo: str,
) -> list[dict[str, object]]:
    if not plan_actual:
        return []

    recipe_ids = list(plan_actual.items.values_list("receta_id", flat=True))
    if not recipe_ids:
        return []

    month = None
    try:
        month = int((periodo or "").split("-")[1])
    except (IndexError, TypeError, ValueError):
        month = None

    try:
        sales_qs = VentaHistorica.objects.filter(receta_id__in=recipe_ids, sucursal_id__isnull=False)
        if month:
            monthly_qs = sales_qs.filter(fecha__month=month)
            sales_qs = monthly_qs if monthly_qs.exists() else sales_qs
        sales_rows = list(
            sales_qs.values("sucursal_id", "sucursal__codigo", "sucursal__nombre")
            .annotate(
                total_units=Sum("cantidad"),
                total_tickets=Sum("tickets"),
                recipe_count=Count("receta_id", distinct=True),
            )
            .order_by("-total_units", "sucursal__codigo")
        )
        request_rows = list(
            SolicitudVenta.objects.filter(
                periodo=periodo,
                receta_id__in=recipe_ids,
                sucursal_id__isnull=False,
            )
            .values("sucursal_id")
            .annotate(
                solicitud_total=Sum("cantidad"),
                solicitud_rows=Count("id"),
            )
        )
    except (OperationalError, ProgrammingError):
        return []

    request_map = {int(row["sucursal_id"]): row for row in request_rows if row.get("sucursal_id")}
    recipe_plan_map = {
        int(item.receta_id): _to_decimal_safe(item.cantidad)
        for item in plan_actual.items.select_related("receta").all()
    }
    branch_recipe_rows = list(
        sales_qs.values("sucursal_id", "receta_id", "receta__nombre")
        .annotate(total_units=Sum("cantidad"))
        .order_by("sucursal_id", "-total_units", "receta__nombre")
    )
    dominant_recipe_map: dict[int, dict[str, object]] = {}
    for recipe_row in branch_recipe_rows:
        sucursal_id = int(recipe_row.get("sucursal_id") or 0)
        receta_id = int(recipe_row.get("receta_id") or 0)
        if not sucursal_id or not receta_id or sucursal_id in dominant_recipe_map:
            continue
        dominant_recipe_map[sucursal_id] = {
            "recipe_name": recipe_row.get("receta__nombre") or "Producto",
            "recipe_units": _to_decimal_safe(recipe_row.get("total_units")),
            "plan_qty": recipe_plan_map.get(receta_id, Decimal("0")),
        }

    rows: list[dict[str, object]] = []
    for row in sales_rows:
        sucursal_id = int(row.get("sucursal_id") or 0)
        if not sucursal_id:
            continue
        historico_units = _to_decimal_safe(row.get("total_units"))
        solicitud = request_map.get(sucursal_id) or {}
        solicitud_total = _to_decimal_safe(solicitud.get("solicitud_total"))
        solicitud_rows = int(solicitud.get("solicitud_rows") or 0)
        recipe_count = int(row.get("recipe_count") or 0)
        dominant_recipe = dominant_recipe_map.get(sucursal_id) or {}
        dominant_recipe_name = str(dominant_recipe.get("recipe_name") or "Producto")
        dominant_recipe_units = _to_decimal_safe(dominant_recipe.get("recipe_units"))
        dominant_plan_qty = _to_decimal_safe(dominant_recipe.get("plan_qty"))

        if solicitud_total > 0 and historico_units <= 0:
            tone = "danger"
            status = "Solicitud sin base"
            detail = "La sucursal ya trae pedido, pero el plan no tiene respaldo comparable para ese periodo."
            priority_score = solicitud_total * Decimal("10")
            action_label = "Revisar solicitud"
        elif solicitud_total > historico_units and historico_units > 0:
            tone = "warning"
            status = "Presión superior al histórico"
            detail = (
                f"Solicitud { _quantize_qty(solicitud_total) } vs histórico comparable "
                f"{ _quantize_qty(historico_units) }."
            )
            priority_score = (solicitud_total * Decimal("8")) + historico_units
            action_label = "Alinear demanda"
        elif solicitud_total > 0:
            tone = "primary"
            status = "Solicitud activa"
            detail = (
                f"La sucursal ya empuja { _quantize_qty(solicitud_total) } unidades para "
                f"{recipe_count} producto(s) del plan."
            )
            priority_score = (solicitud_total * Decimal("6")) + historico_units
            action_label = "Abrir demanda"
        else:
            tone = "success"
            status = "Demanda comparable"
            detail = (
                f"Trae { _quantize_qty(historico_units) } unidades comparables para "
                f"{recipe_count} producto(s) del plan."
            )
            priority_score = historico_units
            action_label = "Ver histórico"

        rows.append(
            {
                "sucursal_codigo": row.get("sucursal__codigo") or "",
                "sucursal_nombre": row.get("sucursal__nombre") or "Sucursal",
                "status": status,
                "tone": tone,
                "detail": detail,
                "historico_units": _quantize_qty(historico_units),
                "solicitud_total": _quantize_qty(solicitud_total),
                "solicitud_rows": solicitud_rows,
                "recipe_count": recipe_count,
                "dominant_recipe_name": dominant_recipe_name,
                "dominant_recipe_units": _quantize_qty(dominant_recipe_units),
                "dominant_plan_qty": _quantize_qty(dominant_plan_qty),
                "action_url": f"{reverse('recetas:plan_produccion')}?plan_id={plan_actual.id}&periodo={periodo}#plan-productos",
                "action_label": action_label,
                "priority_score": priority_score,
            }
        )

    tone_order = {"danger": 0, "warning": 1, "primary": 2, "success": 3}
    rows.sort(
        key=lambda item: (
            tone_order.get(str(item.get("tone") or ""), 9),
            -float(item.get("priority_score") or 0),
            str(item.get("sucursal_codigo") or ""),
        )
    )
    return rows[:6]


def _plan_branch_supply_rows(
    *,
    branch_priority_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    if not branch_priority_rows:
        return []

    recipe_name_map = {
        str(row.get("dominant_recipe_name") or ""): _quantize_qty(_to_decimal_safe(row.get("dominant_plan_qty")))
        for row in branch_priority_rows
    }
    recipe_ids = list(
        Receta.objects.filter(nombre__in=[name for name in recipe_name_map.keys() if name]).values_list("id", flat=True)
    )
    if not recipe_ids:
        return []

    lineas = list(
        LineaReceta.objects.filter(receta_id__in=recipe_ids, insumo_id__isnull=False)
        .exclude(tipo_linea=LineaReceta.TIPO_SUBSECCION)
        .select_related("insumo", "receta", "insumo__unidad_base")
    )
    if not lineas:
        return []

    canonical_by_line: dict[int, Insumo] = {}
    canonical_ids: set[int] = set()
    lineas_by_recipe_name: dict[str, list[LineaReceta]] = defaultdict(list)
    for linea in lineas:
        if not linea.insumo:
            continue
        canonical = canonical_insumo(linea.insumo) or linea.insumo
        canonical_by_line[linea.id] = canonical
        canonical_ids.add(canonical.id)
        lineas_by_recipe_name[linea.receta.nombre].append(linea)

    existencia_map = {
        int(existencia.insumo_id): existencia
        for existencia in ExistenciaInsumo.objects.filter(insumo_id__in=canonical_ids).select_related("insumo")
    }

    rows: list[dict[str, object]] = []
    for branch_row in branch_priority_rows:
        recipe_name = str(branch_row.get("dominant_recipe_name") or "")
        plan_qty = _to_decimal_safe(branch_row.get("dominant_plan_qty"))
        if not recipe_name or plan_qty <= 0:
            continue

        best_candidate: dict[str, object] | None = None
        best_score = Decimal("-1")
        for linea in lineas_by_recipe_name.get(recipe_name, []):
            canonical = canonical_by_line.get(linea.id)
            if canonical is None:
                continue
            required_qty = _to_decimal_safe(linea.cantidad) * plan_qty
            if required_qty <= 0:
                continue
            existencia = existencia_map.get(canonical.id)
            stock_actual = _to_decimal_safe(getattr(existencia, "stock_actual", 0))
            shortage = max(required_qty - stock_actual, Decimal("0"))
            readiness = _insumo_erp_readiness(canonical)
            missing = list(readiness.get("missing") or [])
            missing_cost = _latest_cost_for_insumo(canonical) is None
            score = (shortage * Decimal("100")) + (Decimal(str(len(missing))) * Decimal("50")) + required_qty
            if missing_cost:
                score += Decimal("25")
            if score > best_score:
                best_score = score
                best_candidate = {
                    "insumo_nombre": canonical.nombre,
                    "required_qty": _quantize_qty(required_qty),
                    "stock_actual": _quantize_qty(stock_actual),
                    "shortage": _quantize_qty(shortage),
                    "master_missing": missing,
                    "missing_cost": missing_cost,
                    "unidad": canonical.unidad_base.codigo if getattr(canonical, "unidad_base_id", None) else "-",
                    "action_url": reverse("maestros:insumo_update", args=[canonical.id]),
                    "action_label": "Asegurar insumo",
                }

        if best_candidate:
            rows.append(
                {
                    "sucursal_codigo": branch_row.get("sucursal_codigo") or "",
                    "sucursal_nombre": branch_row.get("sucursal_nombre") or "Sucursal",
                    "dominant_recipe_name": recipe_name,
                    **best_candidate,
                    "priority_score": best_score,
                }
            )

    rows.sort(
        key=lambda item: (
            Decimal(str(item.get("shortage") or 0)),
            Decimal(str(len(item.get("master_missing") or []))),
            Decimal(str(item.get("required_qty") or 0)),
        ),
        reverse=True,
    )
    return rows[:6]


def _plan_status_dashboard(
    plans_qs,
    *,
    start_date: date | None = None,
    end_date: date | None = None,
    group_by: str = "day",
    limit: int = 7,
) -> dict[str, Any]:
    if start_date:
        plans_qs = plans_qs.filter(fecha_produccion__gte=start_date)
    if end_date:
        plans_qs = plans_qs.filter(fecha_produccion__lte=end_date)
    if group_by not in {"day", "week", "month"}:
        group_by = "day"

    total = plans_qs.count()
    borrador = plans_qs.filter(estado=PlanProduccion.ESTADO_BORRADOR).count()
    consumo_aplicado = plans_qs.filter(estado=PlanProduccion.ESTADO_CONSUMO_APLICADO).count()
    cerrados = plans_qs.filter(estado=PlanProduccion.ESTADO_CERRADO).count()
    abiertos = borrador + consumo_aplicado
    oldest_open = plans_qs.exclude(estado=PlanProduccion.ESTADO_CERRADO).order_by("fecha_produccion", "id").first()
    oldest_open_days = 0
    if oldest_open and oldest_open.fecha_produccion:
        oldest_open_days = max((timezone.localdate() - oldest_open.fecha_produccion).days, 0)

    grouped_rows: dict[date, dict[str, Any]] = {}
    for fecha_produccion, estado in plans_qs.values_list("fecha_produccion", "estado"):
        if group_by == "week":
            bucket_date = fecha_produccion - timedelta(days=fecha_produccion.weekday())
            label = f"Semana {bucket_date.isoformat()}"
        elif group_by == "month":
            bucket_date = date(fecha_produccion.year, fecha_produccion.month, 1)
            label = bucket_date.strftime("%Y-%m")
        else:
            bucket_date = fecha_produccion
            label = bucket_date.isoformat()
        bucket = grouped_rows.setdefault(
            bucket_date,
            {
                "group_date": bucket_date,
                "label": label,
                "total": 0,
                "borrador": 0,
                "consumo_aplicado": 0,
                "cerrado": 0,
                "abiertos": 0,
            },
        )
        bucket["total"] += 1
        if estado == PlanProduccion.ESTADO_BORRADOR:
            bucket["borrador"] += 1
        elif estado == PlanProduccion.ESTADO_CONSUMO_APLICADO:
            bucket["consumo_aplicado"] += 1
        elif estado == PlanProduccion.ESTADO_CERRADO:
            bucket["cerrado"] += 1
        bucket["abiertos"] = bucket["borrador"] + bucket["consumo_aplicado"]

    rows = sorted(grouped_rows.values(), key=lambda item: item["group_date"], reverse=True)[:limit]

    if abiertos == 0:
        status = "Operacion al dia"
        tone = "success"
        detail = "No hay planes abiertos; todos los planes registrados ya quedaron cerrados."
    elif consumo_aplicado > 0:
        status = "Pendientes de cierre"
        tone = "warning"
        detail = f"Hay {consumo_aplicado} plan(es) con consumo aplicado pendientes de cierre formal."
    else:
        status = "Pendientes de ejecucion"
        tone = "warning"
        detail = f"Hay {borrador} plan(es) en borrador pendientes de ejecutar."

    return {
        "total": total,
        "borrador": borrador,
        "consumo_aplicado": consumo_aplicado,
        "cerrados": cerrados,
        "abiertos": abiertos,
        "oldest_open_name": getattr(oldest_open, "nombre", ""),
        "oldest_open_date": getattr(oldest_open, "fecha_produccion", None),
        "oldest_open_days": oldest_open_days,
        "status": status,
        "tone": tone,
        "detail": detail,
        "start_date": start_date,
        "end_date": end_date,
        "group_by": group_by,
        "group_by_label": {
            "day": "Fecha producción",
            "week": "Semana",
            "month": "Mes",
        }[group_by],
        "rows": rows,
    }


def _plan_status_dashboard_filters(request: HttpRequest) -> dict[str, Any]:
    start_date = _parse_date_safe(request.GET.get("dg_start_date"))
    end_date = _parse_date_safe(request.GET.get("dg_end_date"))
    group_by = (request.GET.get("dg_group_by") or "day").strip().lower()
    if group_by not in {"day", "week", "month"}:
        group_by = "day"
    return {
        "start_date": start_date,
        "end_date": end_date,
        "group_by": group_by,
    }


def _critical_path_rows_from_cards(
    cards: list[dict[str, object]],
    *,
    owner: str,
    fallback_url: str,
    default_dependency: str = "Inicio del flujo",
) -> list[dict[str, object]]:
    tone_completion = {
        "danger": 20,
        "warning": 55,
        "primary": 75,
        "success": 100,
    }
    stage_rows: list[dict[str, object]] = []
    for card in cards:
        tone = str(card.get("tone") or ("danger" if card.get("count") else "success"))
        count = card.get("count") or 0
        stage_rows.append(
            {
                "label": card.get("label") or "Tramo operativo",
                "owner": owner,
                "open_count": count,
                "progress_pct": tone_completion.get(tone, 0 if count else 100),
                "semaphore_tone": tone,
                "semaphore_label": "En revisión" if count else "Cerrado",
                "detail": card.get("detail") or "Sin detalle operativo",
                "depends_on": default_dependency,
                "dependency_status": card.get("detail") or "Sin dependencia registrada",
                "next_step": card.get("action_detail") or card.get("action_label") or "Revisar tramo",
                "action_label": card.get("action_label") or "Abrir",
                "action_url": card.get("action_url") or fallback_url,
            }
        )
    return _recipes_critical_path_rows(stage_rows, owner=owner, fallback_url=fallback_url)


@login_required
def recetas_list(request: HttpRequest) -> HttpResponse:
    vista = (request.GET.get("vista") or "").strip().lower()
    q = request.GET.get("q", "").strip()
    estado = request.GET.get("estado", "").strip().lower()
    health_status = request.GET.get("health_status", "").strip().lower()
    chain_status = request.GET.get("chain_status", "").strip().lower()
    chain_checkpoint = request.GET.get("chain_checkpoint", "").strip().lower()
    governance_issue = request.GET.get("governance_issue", "").strip().lower()
    enterprise_stage_filter = request.GET.get("enterprise_stage", "").strip().lower()
    tipo = request.GET.get("tipo", "").strip().upper()
    modo_operativo = request.GET.get("modo_operativo", "").strip().upper()
    familia = request.GET.get("familia", "").strip()
    categoria = request.GET.get("categoria", "").strip()

    familias_db = list(
        Receta.objects.exclude(familia__exact="")
        .values_list("familia", flat=True)
        .distinct()
        .order_by("familia")
    )
    familias_catalogo = familias_producto_catalogo(familias_db)
    categorias_catalogo = list(
        Receta.objects.exclude(categoria__exact="")
        .values_list("categoria", flat=True)
        .distinct()
        .order_by("categoria")
    )
    categorias_catalogo = list(
        Receta.objects.exclude(categoria__exact="")
        .values_list("categoria", flat=True)
        .distinct()
        .order_by("categoria")
    )
    familias_top = list(
        Receta.objects.exclude(familia__exact="")
        .values("familia")
        .annotate(total=Count("id"))
        .order_by("-total", "familia")[:8]
    )
    categorias_top = list(
        Receta.objects.exclude(categoria__exact="")
        .values("categoria")
        .annotate(total=Count("id"))
        .order_by("-total", "categoria")[:8]
    )

    recetas_base = Receta.objects.all()
    total_all = recetas_base.count()
    total_productos = recetas_base.filter(tipo=Receta.TIPO_PRODUCTO_FINAL).count()
    total_insumos = recetas_base.filter(tipo=Receta.TIPO_PREPARACION).count()
    total_subinsumos = recetas_base.filter(
        tipo=Receta.TIPO_PREPARACION,
        usa_presentaciones=True,
    ).count()
    total_batidas_base = max(total_insumos - total_subinsumos, 0)

    # Vista rápida por botones: por defecto mostramos Productos para evitar lista mezclada.
    if vista not in {"productos", "insumos", "subinsumos", "todo"}:
        if tipo == Receta.TIPO_PRODUCTO_FINAL:
            vista = "productos"
        elif tipo == Receta.TIPO_PREPARACION:
            vista = "insumos"
        else:
            vista = "productos"

    recetas = recetas_base.select_related("rendimiento_unidad").annotate(
        pendientes_count=Count(
            "lineas",
            filter=Q(lineas__match_status=LineaReceta.STATUS_NEEDS_REVIEW),
        ),
        lineas_count=Count("lineas"),
        presentaciones_activas_count=Count("presentaciones", filter=Q(presentaciones__activo=True)),
    )
    if vista == "productos":
        recetas = recetas.filter(tipo=Receta.TIPO_PRODUCTO_FINAL)
    elif vista == "insumos":
        recetas = recetas.filter(tipo=Receta.TIPO_PREPARACION)
    elif vista == "subinsumos":
        recetas = recetas.filter(tipo=Receta.TIPO_PREPARACION, usa_presentaciones=True)

    if modo_operativo == "BASE":
        if isinstance(recetas, list):
            recetas = [r for r in recetas if r.tipo == Receta.TIPO_PREPARACION and not r.usa_presentaciones]
        else:
            recetas = recetas.filter(tipo=Receta.TIPO_PREPARACION, usa_presentaciones=False)
    elif modo_operativo == "BASE_DERIVADOS":
        if isinstance(recetas, list):
            recetas = [r for r in recetas if r.tipo == Receta.TIPO_PREPARACION and r.usa_presentaciones]
        else:
            recetas = recetas.filter(tipo=Receta.TIPO_PREPARACION, usa_presentaciones=True)
    elif modo_operativo == "FINAL":
        if isinstance(recetas, list):
            recetas = [r for r in recetas if r.tipo == Receta.TIPO_PRODUCTO_FINAL]
        else:
            recetas = recetas.filter(tipo=Receta.TIPO_PRODUCTO_FINAL)

    if q:
        query_norm = normalizar_nombre(q)
        recetas_ranked: list[tuple[float, Receta]] = []
        for receta in recetas:
            score = _recipe_search_score(query_norm, receta)
            if score >= 60.0:
                recetas_ranked.append((score, receta))
        recetas_ranked.sort(key=lambda item: (-item[0], normalizar_nombre(item[1].nombre)))
        recetas = [item[1] for item in recetas_ranked]
    if tipo in {Receta.TIPO_PREPARACION, Receta.TIPO_PRODUCTO_FINAL}:
        if isinstance(recetas, list):
            recetas = [r for r in recetas if r.tipo == tipo]
        else:
            recetas = recetas.filter(tipo=tipo)
    if familia:
        if isinstance(recetas, list):
            recetas = [r for r in recetas if (r.familia or "") == familia]
        else:
            recetas = recetas.filter(familia=familia)
    if categoria:
        if isinstance(recetas, list):
            recetas = [r for r in recetas if (r.categoria or "") == categoria]
        else:
            recetas = recetas.filter(categoria=categoria)
    if estado == "pendientes":
        if isinstance(recetas, list):
            recetas = [r for r in recetas if (r.pendientes_count or 0) > 0]
        else:
            recetas = recetas.filter(pendientes_count__gt=0)
    elif estado == "ok":
        if isinstance(recetas, list):
            recetas = [r for r in recetas if (r.pendientes_count or 0) == 0]
        else:
            recetas = recetas.filter(pendientes_count=0)
    if health_status in {"listas", "pendientes", "incompletas"}:
        if isinstance(recetas, list):
            recetas = [r for r in recetas if _matches_recipe_health_filter(r, health_status)]
        else:
            recetas = [r for r in recetas if _matches_recipe_health_filter(r, health_status)]
    if chain_status in {"listas", "pendientes", "incompletas"}:
        if isinstance(recetas, list):
            recetas = [
                r for r in recetas
                if (
                    (_recipe_chain_status(r)["code"] == "success" and chain_status == "listas")
                    or (_recipe_chain_status(r)["code"] == "warning" and chain_status == "pendientes")
                    or (_recipe_chain_status(r)["code"] == "danger" and chain_status == "incompletas")
                )
            ]
        else:
            recetas = [
                r for r in recetas
                if (
                    (_recipe_chain_status(r)["code"] == "success" and chain_status == "listas")
                    or (_recipe_chain_status(r)["code"] == "warning" and chain_status == "pendientes")
                    or (_recipe_chain_status(r)["code"] == "danger" and chain_status == "incompletas")
                )
            ]
    if chain_checkpoint in {"base_ready", "derived_sync", "final_usage", "upstream_trace", "packaging_ready", "internal_components"}:
        if isinstance(recetas, list):
            recetas = [r for r in recetas if chain_checkpoint in _recipe_failing_chain_checkpoint_codes(r)]
        else:
            recetas = [r for r in recetas if chain_checkpoint in _recipe_failing_chain_checkpoint_codes(r)]
    if governance_issue in {"familia", "categoria", "maestro_incompleto", "rendimiento", "derivados", "sync_derivados", "sin_consumo_final", "sin_base_origen", "sin_empaque", "base_directa", "componentes"}:
        if isinstance(recetas, list):
            recetas = [r for r in recetas if governance_issue in _recipe_governance_issues(r)]
        else:
            recetas = [r for r in recetas if governance_issue in _recipe_governance_issues(r)]
    if enterprise_stage_filter in {
        "base_setup",
        "derivados_setup",
        "ready_for_final",
        "feeding_final",
        "base_operativa",
        "final_setup",
        "final_normalization",
        "final_packaging",
        "final_ready",
        "final_defined",
    }:
        if isinstance(recetas, list):
            recetas = [r for r in recetas if _matches_recipe_enterprise_stage(r, enterprise_stage_filter)]
        else:
            recetas = [r for r in recetas if _matches_recipe_enterprise_stage(r, enterprise_stage_filter)]
    if not isinstance(recetas, list):
        recetas = recetas.order_by("nombre")

    if isinstance(recetas, list):
        total_recetas = len(recetas)
        total_pendientes = sum(1 for r in recetas if (r.pendientes_count or 0) > 0)
        total_lineas = sum((r.lineas_count or 0) for r in recetas)
    else:
        total_recetas = recetas.count()
        total_pendientes = recetas.filter(pendientes_count__gt=0).count()
        total_lineas = sum(r.lineas_count for r in recetas)

    qs_filters = {
        "vista": vista,
        "q": q,
        "estado": estado,
        "health_status": health_status,
        "chain_status": chain_status,
        "chain_checkpoint": chain_checkpoint,
        "governance_issue": governance_issue,
        "enterprise_stage": enterprise_stage_filter,
        "tipo": tipo,
        "modo_operativo": modo_operativo,
        "familia": familia,
        "categoria": categoria,
    }
    qs_base = urlencode({k: v for k, v in qs_filters.items() if v})

    paginator = Paginator(recetas, 20)
    page = paginator.get_page(request.GET.get("page"))
    for receta in page.object_list:
        receta.operational_health = _recipe_operational_health(receta)
        receta.derived_state = _recipe_derived_sync_state(receta) if receta.tipo == Receta.TIPO_PREPARACION else None
        receta.supply_chain_snapshot = _recipe_supply_chain_snapshot(receta)
        receta.product_upstream_snapshot = None
        receta.direct_base_snapshot = {
            "count": 0,
            "base_names": [],
            "suggested_count": 0,
            "exact_count": 0,
            "sample_suggestions": [],
        }
        if receta.tipo == Receta.TIPO_PRODUCTO_FINAL:
            receta.product_upstream_snapshot = getattr(receta, "_product_upstream_snapshot_cache", None)
            if receta.product_upstream_snapshot is None:
                lineas_qs = list(
                    receta.lineas.select_related("insumo").only(
                        "id",
                        "insumo_id",
                        "insumo__id",
                        "insumo__nombre",
                        "insumo__tipo_item",
                        "insumo__codigo",
                    )
                )
                receta.product_upstream_snapshot = _product_upstream_snapshot(lineas_qs, receta=receta)
                setattr(receta, "_product_upstream_snapshot_cache", receta.product_upstream_snapshot)
            receta.direct_base_snapshot = _recipe_direct_base_snapshot(receta)
        receta.governance_issues = _recipe_governance_issues(receta)
        receta.master_gap_summary = _recipe_master_gap_summary(receta)
        receta.primary_action = _recipe_primary_action(receta)
    health_summary = {"listas": 0, "pendientes": 0, "incompletas": 0}
    chain_summary = {"listas": 0, "pendientes": 0, "incompletas": 0}
    checkpoint_summary = {
        "base_ready": 0,
        "derived_sync": 0,
        "final_usage": 0,
        "upstream_trace": 0,
        "packaging_ready": 0,
        "internal_components": 0,
    }
    source_for_health = recetas if isinstance(recetas, list) else list(recetas)
    governance_summary = {
        "familia": 0,
        "categoria": 0,
        "maestro_incompleto": 0,
        "rendimiento": 0,
        "derivados": 0,
        "sync_derivados": 0,
        "sin_consumo_final": 0,
        "sin_base_origen": 0,
        "sin_empaque": 0,
        "base_directa": 0,
        "componentes": 0,
    }
    master_gap_totals = {
        "unidad": 0,
        "proveedor": 0,
        "categoria": 0,
        "codigo_point": 0,
        "inactivo": 0,
    }
    direct_base_suggested_total = 0
    enterprise_stage_summary = {
        "base_setup": 0,
        "derivados_setup": 0,
        "ready_for_final": 0,
        "feeding_final": 0,
        "base_operativa": 0,
        "final_setup": 0,
        "final_normalization": 0,
        "final_packaging": 0,
        "final_ready": 0,
        "final_defined": 0,
    }
    for receta in source_for_health:
        health = getattr(receta, "operational_health", None) or _recipe_operational_health(receta)
        if health["code"] == "success":
            health_summary["listas"] += 1
        elif health["code"] == "warning":
            health_summary["pendientes"] += 1
        else:
            health_summary["incompletas"] += 1
        issues = getattr(receta, "governance_issues", None) or _recipe_governance_issues(receta)
        chain = getattr(receta, "chain_status_info", None) or _recipe_chain_status(receta)
        if chain["code"] == "success":
            chain_summary["listas"] += 1
        elif chain["code"] == "warning":
            chain_summary["pendientes"] += 1
        else:
            chain_summary["incompletas"] += 1
        for checkpoint_code in _recipe_failing_chain_checkpoint_codes(receta):
            if checkpoint_code in checkpoint_summary:
                checkpoint_summary[checkpoint_code] += 1
        for issue in issues:
            governance_summary[issue] += 1
        gap_summary = getattr(receta, "master_gap_summary", None) or _recipe_master_gap_summary(receta)
        for gap_key, gap_count in gap_summary["counts"].items():
            if gap_key in master_gap_totals:
                master_gap_totals[gap_key] += int(gap_count or 0)
        if receta.tipo == Receta.TIPO_PRODUCTO_FINAL:
            direct_base_snapshot = getattr(receta, "direct_base_snapshot", None) or _recipe_direct_base_snapshot(receta)
            direct_base_suggested_total += int(direct_base_snapshot.get("suggested_count") or 0)
        stage = getattr(receta, "enterprise_stage", None) or _recipe_enterprise_stage(receta)
        enterprise_stage_summary[stage["code"]] += 1
    chain_focus = _build_recipe_chain_focus(
        vista=vista,
        chain_summary=chain_summary,
        checkpoint_summary=checkpoint_summary,
        governance_summary=governance_summary,
        filters=qs_filters,
    )
    erp_governance_rows = _recipe_catalog_governance_rows(
        total_pendientes=total_pendientes,
        health_summary=health_summary,
        master_gap_totals=master_gap_totals,
        chain_focus=chain_focus,
    )
    catalog_master_blockers = int(sum(master_gap_totals.values()))
    catalog_chain_blockers = int(chain_summary["pendientes"] + chain_summary["incompletas"])
    catalog_bom_blockers = int(health_summary["pendientes"] + health_summary["incompletas"])
    downstream_handoff_rows = [
        {
            "label": "MRP",
            "owner": "Planeación / Producción",
            "status": "Listo"
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and catalog_bom_blockers == 0
            else "Bloqueado",
            "tone": "success"
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and catalog_bom_blockers == 0
            else "warning",
            "blockers": catalog_master_blockers + catalog_chain_blockers + catalog_bom_blockers,
            "completion": 100
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and catalog_bom_blockers == 0
            else max(0, 100 - ((catalog_master_blockers + catalog_chain_blockers + catalog_bom_blockers) * 3)),
            "depends_on": "Maestro + BOM + cadena cerrados",
            "exit_criteria": "El catálogo debe explotar demanda sin referencias abiertas ni documentos incompletos.",
            "detail": (
                "El catálogo ya puede alimentar el cálculo de planeación y explosión de demanda."
                if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and catalog_bom_blockers == 0
                else "MRP sigue condicionado por artículos incompletos, cadena abierta o documentos en revisión."
            ),
            "next_step": "Abrir MRP" if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and catalog_bom_blockers == 0 else "Cerrar brechas del catálogo",
            "url": reverse("recetas:mrp_form") if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and catalog_bom_blockers == 0 else (chain_focus.get("action_url") or (reverse("recetas:recetas_list") + (f'?{qs_base}' if qs_base else ""))),
            "cta": "Abrir MRP" if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and catalog_bom_blockers == 0 else (chain_focus.get("action_label") or "Abrir catálogo"),
        },
        {
            "label": "Compras",
            "owner": "Compras / Planeación",
            "status": "Listo"
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
            else "Bloqueado",
            "tone": "success"
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
            else "warning",
            "blockers": catalog_master_blockers + catalog_chain_blockers + int(governance_summary["sin_empaque"]),
            "completion": 100
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
            else max(0, 100 - ((catalog_master_blockers + catalog_chain_blockers + int(governance_summary["sin_empaque"])) * 3)),
            "depends_on": "Empaque + cadena documental",
            "exit_criteria": "Solicitudes y órdenes deben operar con productos finales completos y empaques definidos.",
            "detail": (
                "Compras ya puede usar el catálogo como referencia documental de abastecimiento."
                if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
                else "Compras sigue bloqueado por empaques faltantes, cadena abierta o artículos sin cierre maestro."
            ),
            "next_step": "Abrir compras" if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0 else "Cerrar empaque y maestro",
            "url": reverse("compras:solicitudes") if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0 else (reverse("recetas:recetas_list") + "?vista=productos&governance_issue=sin_empaque"),
            "cta": "Abrir compras" if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0 else "Ver productos finales",
        },
        {
            "label": "Inventario",
            "owner": "Inventario / Almacén",
            "status": "Listo"
            if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
            else "Bloqueado",
            "tone": "success"
            if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
            else "warning",
            "blockers": catalog_master_blockers + int(checkpoint_summary["packaging_ready"]),
            "completion": 100
            if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
            else max(0, 100 - ((catalog_master_blockers + int(checkpoint_summary["packaging_ready"])) * 4)),
            "depends_on": "Artículo maestro + empaque",
            "exit_criteria": "Inventario debe recibir productos y empaques como artículos operativos estables.",
            "detail": (
                "Inventario ya puede trabajar el catálogo sin artículos abiertos del maestro."
                if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
                else "Inventario todavía requiere artículos completos del maestro o empaques cerrados."
            ),
            "next_step": "Abrir inventario" if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0 else "Cerrar maestro e inventario",
            "url": reverse("inventario:existencias") if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0 else reverse("maestros:insumo_list"),
            "cta": "Abrir inventario" if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0 else "Abrir maestro",
        },
    ]
    trunk_handoff_rows = [
        {
            "label": "Recetas / BOM",
            "owner": "Recetas / Producción",
            "status": "Listo para operar"
            if catalog_master_blockers == 0 and catalog_bom_blockers == 0 and catalog_chain_blockers == 0
            else "Bloqueado",
            "tone": "success"
            if catalog_master_blockers == 0 and catalog_bom_blockers == 0 and catalog_chain_blockers == 0
            else "warning",
            "blockers": catalog_master_blockers + catalog_bom_blockers + catalog_chain_blockers,
            "completion": 100
            if catalog_master_blockers == 0 and catalog_bom_blockers == 0 and catalog_chain_blockers == 0
            else max(0, 100 - ((catalog_master_blockers + catalog_bom_blockers + catalog_chain_blockers) * 3)),
            "depends_on": "Maestro cerrado + estructura base-derivados-final consistente",
            "exit_criteria": "Las recetas deben costear, normalizar y alimentar producto final sin referencias abiertas.",
            "detail": (
                "El catálogo ya puede operar como documento BOM estable."
                if catalog_master_blockers == 0 and catalog_bom_blockers == 0 and catalog_chain_blockers == 0
                else "Todavía hay brechas en maestro, cadena o estructura BOM."
            ),
            "next_step": chain_focus.get("action_detail") or "Cerrar catálogo de recetas",
            "url": chain_focus.get("action_url") or (reverse("recetas:recetas_list") + (f'?{qs_base}' if qs_base else "")),
            "cta": chain_focus.get("action_label") or "Abrir recetas",
        },
        {
            "label": "Compras documentales",
            "owner": "Compras / Planeación",
            "status": "Listo para operar"
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
            else "Bloqueado",
            "tone": "success"
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
            else "warning",
            "blockers": catalog_master_blockers + catalog_chain_blockers + int(governance_summary["sin_empaque"]),
            "completion": 100
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
            else max(0, 100 - ((catalog_master_blockers + catalog_chain_blockers + int(governance_summary["sin_empaque"])) * 3)),
            "depends_on": "Producto final completo + empaques definidos",
            "exit_criteria": "Solicitudes, órdenes y recepciones deben usar productos finales cerrados documentalmente.",
            "detail": (
                "Compras ya puede tomar el catálogo como documento operativo de abastecimiento."
                if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
                else "Compras sigue condicionado por maestro, cadena o empaques pendientes."
            ),
            "next_step": "Abrir compras" if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0 else "Cerrar empaque y maestro",
            "url": reverse("compras:solicitudes") if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0 else (reverse("recetas:recetas_list") + "?vista=productos&governance_issue=sin_empaque"),
            "cta": "Abrir compras" if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0 else "Ver productos finales",
        },
        {
            "label": "Inventario / Reabasto",
            "owner": "Inventario / Almacén",
            "status": "Listo para operar"
            if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
            else "Bloqueado",
            "tone": "success"
            if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
            else "warning",
            "blockers": catalog_master_blockers + int(checkpoint_summary["packaging_ready"]),
            "completion": 100
            if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
            else max(0, 100 - ((catalog_master_blockers + int(checkpoint_summary["packaging_ready"])) * 4)),
            "depends_on": "Artículos listos + empaques operativos + cierre documental",
            "exit_criteria": "Inventario debe trabajar finales y empaques como artículos estables de stock y reabasto.",
            "detail": (
                "Inventario ya puede sostener existencia y reabasto sobre el catálogo final."
                if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
                else "Inventario todavía requiere cierre de maestro o empaque para operar estable."
            ),
            "next_step": "Abrir inventario" if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0 else "Cerrar maestro e inventario",
            "url": reverse("inventario:existencias") if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0 else reverse("maestros:insumo_list"),
            "cta": "Abrir inventario" if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0 else "Abrir maestro",
        },
    ]
    for receta in page.object_list:
        receta.chain_status_info = _recipe_chain_status(receta)
        receta.chain_checkpoints = _recipe_chain_checkpoints(receta)
        receta.chain_action_links = _recipe_chain_actions_catalog(receta)
        receta.chain_focus_summary = _recipe_chain_focus_summary(receta)
        receta.enterprise_stage = _recipe_enterprise_stage(receta)
        receta.enterprise_stage_playbook = _recipe_enterprise_stage_playbook(receta)
        receta.enterprise_stage_progress = _recipe_stage_progress(receta.enterprise_stage_playbook)
        receta.document_status = _recipe_document_status(receta)
    return render(
        request,
        "recetas/recetas_list.html",
        {
            "page": page,
            "vista": vista,
            "q": q,
            "estado": estado,
            "health_status": health_status,
            "chain_status": chain_status,
            "chain_checkpoint": chain_checkpoint,
            "governance_issue": governance_issue,
            "tipo": tipo,
            "modo_operativo": modo_operativo,
            "familia": familia,
            "categoria": categoria,
            "familias_catalogo": familias_catalogo,
            "categorias_catalogo": categorias_catalogo,
            "familias_top": familias_top,
            "categorias_top": categorias_top,
            "total_recetas": total_recetas,
            "total_pendientes": total_pendientes,
            "total_lineas": total_lineas,
            "total_all": total_all,
            "total_productos": total_productos,
            "total_insumos": total_insumos,
            "total_batidas_base": total_batidas_base,
            "total_subinsumos": total_subinsumos,
            "health_summary": health_summary,
            "chain_summary": chain_summary,
            "chain_focus": chain_focus,
            "enterprise_stage_filter": enterprise_stage_filter,
            "enterprise_stage_summary": enterprise_stage_summary,
            "checkpoint_summary": checkpoint_summary,
            "governance_summary": governance_summary,
            "master_gap_totals": master_gap_totals,
            "erp_command_center": {
                "owner": "Recetas / Producción",
                "status": "Crítico"
                if (health_summary["pendientes"] + health_summary["incompletas"] + sum(master_gap_totals.values())) > 0
                else "Controlado",
                "tone": "danger"
                if (health_summary["pendientes"] + health_summary["incompletas"] + sum(master_gap_totals.values())) > 0
                else "success",
                "blockers": health_summary["pendientes"] + health_summary["incompletas"] + sum(master_gap_totals.values()),
                "next_step": chain_focus.get("action_detail") or "Continuar el cierre operativo del catálogo de recetas.",
                "url": chain_focus.get("action_url") or qs_base,
                "cta": chain_focus.get("action_label") or "Abrir catálogo",
            },
            "erp_governance_rows": erp_governance_rows,
            "downstream_handoff_rows": downstream_handoff_rows,
            "trunk_handoff_rows": trunk_handoff_rows,
            "trunk_handoff_summary": _trunk_handoff_summary(
                trunk_handoff_rows,
                owner="Recetas / Producción",
                fallback_url=qs_base,
            ),
            "executive_radar_rows": _recipes_executive_radar_rows(
                erp_governance_rows,
                owner="Recetas / Producción",
                fallback_url=qs_base,
            ),
            "critical_path_rows": _recipes_critical_path_rows(
                erp_governance_rows,
                owner="Recetas / Producción",
                fallback_url=qs_base,
            ),
            "direct_base_suggested_total": direct_base_suggested_total,
            "qs_base": qs_base,
        },
    )


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def recetas_sync_all(request: HttpRequest) -> HttpResponse:
    next_url = request.POST.get("next") or reverse("recetas:recetas_list")
    try:
        job, summary = _run_point_recipe_sync_action(request, action_label="SYNC_ALL_RECIPES")
        snapshot = _snapshot_costeo_after_sync()
        messages.success(
            request,
            (
                f"Point recetas actualizado: {summary.get('products_selected', 0)} productos revisados, "
                f"{summary.get('lineas_created', 0)} líneas materializadas. "
                f"Corte semanal {snapshot['week_start']} recalculado."
            ),
        )
        log_event(
            request.user,
            "SNAPSHOT_WEEKLY_COST",
            "recetas.RecetaCostoSemanal",
            None,
            {"source": "SYNC_ALL_RECIPES", "job_id": job.id, "snapshot": snapshot},
        )
    except Exception as exc:
        messages.error(request, f"No se pudo actualizar recetas desde Point: {exc}")
    return redirect(next_url)


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def recetas_sync_group(request: HttpRequest) -> HttpResponse:
    next_url = request.POST.get("next") or reverse("recetas:recetas_list")
    qs = _filtered_recipe_sync_queryset(request.POST).exclude(codigo_point__exact="")
    product_codes = list(qs.values_list("codigo_point", flat=True).distinct())
    if not product_codes:
        messages.warning(request, "No hay recetas filtradas con código Point para sincronizar.")
        return redirect(next_url)
    try:
        job, summary = _run_point_recipe_sync_action(
            request,
            action_label="SYNC_FILTERED_GROUP",
            product_codes=product_codes,
        )
        receta_ids = list(Receta.objects.filter(codigo_point__in=product_codes).values_list("id", flat=True))
        snapshot = _snapshot_costeo_after_sync(receta_ids=receta_ids)
        messages.success(
            request,
            (
                f"Grupo actualizado desde Point: {len(product_codes)} códigos enviados, "
                f"{summary.get('products_selected', 0)} productos procesados. "
                f"Snapshot semanal del grupo recalculado."
            ),
        )
        log_event(
            request.user,
            "SNAPSHOT_WEEKLY_COST",
            "recetas.RecetaCostoSemanal",
            None,
            {"source": "SYNC_FILTERED_GROUP", "job_id": job.id, "snapshot": snapshot, "product_codes": product_codes},
        )
    except Exception as exc:
        messages.error(request, f"No se pudo actualizar el grupo filtrado: {exc}")
    return redirect(next_url)


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def recetas_sync_new(request: HttpRequest) -> HttpResponse:
    next_url = request.POST.get("next") or reverse("recetas:recetas_list")
    service = PointProductRecipeSyncService()
    try:
        discovery = service.discover_new_product_codes(branch_hint="MATRIZ")
        new_codes = discovery.get("new_codes") or []
        if not new_codes:
            messages.info(request, "Point no reportó productos nuevos con receta pendientes de incorporar.")
            return redirect(next_url)
        job, summary = _run_point_recipe_sync_action(
            request,
            action_label="SYNC_ONLY_NEW_PRODUCTS",
            product_codes=list(new_codes),
        )
        receta_ids = list(Receta.objects.filter(codigo_point__in=new_codes).values_list("id", flat=True))
        snapshot = _snapshot_costeo_after_sync(receta_ids=receta_ids)
        messages.success(
            request,
            (
                f"Se incorporaron {len(new_codes)} códigos nuevos desde Point. "
                f"{summary.get('products_selected', 0)} productos procesados y snapshot semanal actualizado."
            ),
        )
        log_event(
            request.user,
            "SYNC_ONLY_NEW_PRODUCTS",
            "pos_bridge.PointSyncJob",
            job.id,
            {"discovery": discovery, "snapshot": snapshot},
        )
    except Exception as exc:
        messages.error(request, f"No se pudo incorporar productos nuevos desde Point: {exc}")
    return redirect(next_url)


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def receta_sync_point(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    next_url = request.POST.get("next") or reverse("recetas:receta_detail", args=[receta.id])
    if not (receta.codigo_point or "").strip():
        messages.warning(request, "La receta no tiene código Point para sincronizar.")
        return redirect(next_url)
    try:
        job, summary = _run_point_recipe_sync_action(
            request,
            action_label="SYNC_SINGLE_RECIPE",
            product_codes=[receta.codigo_point],
        )
        snapshot = _snapshot_costeo_after_sync(receta_ids=[receta.id])
        messages.success(
            request,
            (
                f"Receta {receta.codigo_point} actualizada desde Point. "
                f"{summary.get('lineas_created', 0)} líneas materializadas y costeo semanal recalculado."
            ),
        )
        log_event(
            request.user,
            "SNAPSHOT_WEEKLY_COST",
            "recetas.RecetaCostoSemanal",
            None,
            {"source": "SYNC_SINGLE_RECIPE", "job_id": job.id, "receta_id": receta.id, "snapshot": snapshot},
        )
    except Exception as exc:
        messages.error(request, f"No se pudo actualizar la receta desde Point: {exc}")
    return redirect(next_url)


def _costeo_scope_queryset(selected_week: date, *, scope: str, familia: str = "", temporalidad: str = "", q: str = ""):
    qs = RecetaCostoSemanal.objects.filter(week_start=selected_week).select_related("receta", "base_receta", "addon_receta", "addon_rule")
    if scope == "productos":
        qs = qs.filter(scope_type=RecetaCostoSemanal.SCOPE_RECIPE, receta__tipo=Receta.TIPO_PRODUCTO_FINAL)
    elif scope == "bases":
        qs = qs.filter(scope_type=RecetaCostoSemanal.SCOPE_RECIPE, receta__tipo=Receta.TIPO_PREPARACION)
    elif scope == "addons":
        qs = qs.filter(scope_type=RecetaCostoSemanal.SCOPE_GROUPED_ADDON)
    if familia:
        qs = qs.filter(familia=familia)
    if temporalidad:
        qs = qs.filter(temporalidad=temporalidad)
    if q:
        qs = qs.filter(
            Q(label__icontains=q)
            | Q(base_receta__nombre__icontains=q)
            | Q(addon_receta__nombre__icontains=q)
            | Q(receta__nombre__icontains=q)
        )
    return qs.order_by("label", "id")


@login_required
@permission_required("recetas.view_receta", raise_exception=True)
def costeo_dashboard(request: HttpRequest) -> HttpResponse:
    week_values = list(
        RecetaCostoSemanal.objects.order_by("-week_start").values_list("week_start", flat=True).distinct()[:26]
    )
    selected_week = None
    week_raw = (request.GET.get("week_start") or "").strip()
    if week_raw:
        try:
            selected_week = date.fromisoformat(week_raw)
        except ValueError:
            selected_week = None
    if selected_week is None and week_values:
        selected_week = week_values[0]

    scope = (request.GET.get("scope") or "productos").strip().lower()
    if scope not in {"productos", "bases", "addons", "todo"}:
        scope = "productos"
    familia = (request.GET.get("familia") or "").strip()
    temporalidad = (request.GET.get("temporalidad") or "").strip().upper()
    q = (request.GET.get("q") or "").strip()

    familias_catalogo = list(
        RecetaCostoSemanal.objects.exclude(familia__exact="")
        .values_list("familia", flat=True)
        .distinct()
        .order_by("familia")
    )

    rows = []
    selected_identity = None
    history_rows = []
    if selected_week is not None:
        rows = list(_costeo_scope_queryset(selected_week, scope=scope, familia=familia, temporalidad=temporalidad, q=q))
        selected_identity = (request.GET.get("identity") or "").strip() or (rows[0].identity_key if rows else "")
        if selected_identity:
            history_rows = list(
                RecetaCostoSemanal.objects.filter(identity_key=selected_identity)
                .select_related("receta", "base_receta", "addon_receta")
                .order_by("-week_start", "-id")[:12]
            )
    total_cost = sum(Decimal(row.costo_total or 0) for row in rows)
    total_delta = sum(Decimal(row.delta_total or 0) for row in rows if row.delta_total is not None)
    increase_rows = sorted(
        [row for row in rows if row.delta_total is not None and Decimal(row.delta_total) > 0],
        key=lambda item: Decimal(item.delta_total),
        reverse=True,
    )[:10]
    decrease_rows = sorted(
        [row for row in rows if row.delta_total is not None and Decimal(row.delta_total) < 0],
        key=lambda item: Decimal(item.delta_total),
    )[:10]
    stable_rows = len([row for row in rows if row.delta_total in (None, Decimal("0"), 0)])
    selected_row = next((row for row in rows if row.identity_key == selected_identity), None)
    max_history_cost = max((Decimal(item.costo_total or 0) for item in history_rows), default=Decimal("0"))
    for item in history_rows:
        current_cost = Decimal(item.costo_total or 0)
        item.history_pct = float((current_cost / max_history_cost) * Decimal("100")) if max_history_cost > 0 else 0.0
    previous_week = week_values[1] if len(week_values) > 1 else None

    return render(
        request,
        "recetas/costeo_dashboard.html",
        {
            "selected_week": selected_week,
            "week_values": week_values,
            "previous_week": previous_week,
            "scope": scope,
            "familia": familia,
            "temporalidad": temporalidad,
            "q": q,
            "familias_catalogo": familias_catalogo,
            "rows": rows,
            "increase_rows": increase_rows,
            "decrease_rows": decrease_rows,
            "stable_rows": stable_rows,
            "total_cost": total_cost,
            "total_delta": total_delta,
            "selected_identity": selected_identity,
            "selected_row": selected_row,
            "history_rows": history_rows,
            "max_history_cost": max_history_cost,
        },
    )


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def costeo_dashboard_snapshot(request: HttpRequest) -> HttpResponse:
    next_url = request.POST.get("next") or reverse("recetas:costeo_dashboard")
    try:
        summary = snapshot_weekly_costs()
        messages.success(
            request,
            (
                f"Corte semanal de costeo actualizado: semana {summary.week_start} a {summary.week_end}, "
                f"{summary.total_items} registros procesados."
            ),
        )
        log_event(
            request.user,
            "SNAPSHOT_WEEKLY_COST",
            "recetas.RecetaCostoSemanal",
            None,
            {
                "week_start": summary.week_start.isoformat(),
                "week_end": summary.week_end.isoformat(),
                "recipes_created": summary.recipes_created,
                "recipes_updated": summary.recipes_updated,
                "addons_created": summary.addons_created,
                "addons_updated": summary.addons_updated,
            },
        )
    except Exception as exc:
        messages.error(request, f"No se pudo generar el corte semanal de costeo: {exc}")
    return redirect(next_url)


@login_required
@permission_required("recetas.add_receta", raise_exception=True)
def receta_create(request: HttpRequest) -> HttpResponse:
    familias_db = list(
        Receta.objects.exclude(familia__exact="")
        .values_list("familia", flat=True)
        .distinct()
        .order_by("familia")
    )
    familias_catalogo = familias_producto_catalogo(familias_db)
    categorias_catalogo = list(
        Receta.objects.exclude(categoria__exact="")
        .values_list("categoria", flat=True)
        .distinct()
        .order_by("categoria")
    )
    unidades = UnidadMedida.objects.order_by("codigo")

    recipe_modes = [
        {
            "code": "BASE",
            "title": "Insumo interno base",
            "description": "Batida o mezcla producida con materia prima. Puede llevar rendimiento y luego generar presentaciones.",
        },
        {
            "code": "BASE_DERIVADOS",
            "title": "Insumo base con presentaciones",
            "description": "Usa este modo si de la batida saldrán tamaños o subinsumos derivados como chico, mediano, grande o rosca.",
        },
        {
            "code": "FINAL",
            "title": "Producto final de venta",
            "description": "Artículo final armado con panes, rellenos, coberturas, empaques u otros componentes.",
        },
    ]

    defaults = {
        "nombre": "",
        "codigo_point": "",
        "familia": "",
        "categoria": "",
        "sheet_name": "",
        "recipe_mode": "BASE",
        "rendimiento_cantidad": "",
        "rendimiento_unidad_id": "",
    }

    requested_mode = (request.GET.get("mode") or "").strip().upper()
    if requested_mode in {"BASE", "BASE_DERIVADOS", "FINAL"}:
        defaults["recipe_mode"] = requested_mode
    source_base_receta = None
    source_base_context = None
    source_base_raw = (request.GET.get("source_base") or "").strip()
    if source_base_raw.isdigit():
        source_base_receta = (
            Receta.objects.filter(pk=int(source_base_raw), tipo=Receta.TIPO_PREPARACION)
            .only(
                "id",
                "nombre",
                "familia",
                "categoria",
                "sheet_name",
                "usa_presentaciones",
                "rendimiento_cantidad",
                "rendimiento_unidad_id",
            )
            .first()
        )
        if source_base_receta and defaults["recipe_mode"] == "FINAL":
            if source_base_receta.familia:
                defaults["familia"] = source_base_receta.familia
            defaults["categoria"] = source_base_receta.categoria or defaults["categoria"]
            defaults["sheet_name"] = source_base_receta.sheet_name or defaults["sheet_name"]
            derived_state = _recipe_derived_sync_state(source_base_receta)
            active_presentaciones = int(derived_state["active_presentaciones"] or 0)
            derived_presentaciones = int(derived_state["derived_presentaciones"] or 0)
            source_base_context = {
                "base_ready": bool(
                    source_base_receta.rendimiento_cantidad
                    and source_base_receta.rendimiento_unidad_id
                    and derived_state["prep_ready"]
                ),
                "uses_presentaciones": bool(source_base_receta.usa_presentaciones),
                "active_presentaciones": active_presentaciones,
                "derived_presentaciones": derived_presentaciones,
                "sync_missing_count": max(active_presentaciones - derived_presentaciones, 0),
                "next_step_label": (
                    "La base ya puede alimentar el producto final."
                    if active_presentaciones <= 0 or derived_presentaciones >= active_presentaciones
                    else "Conviene sincronizar derivados antes de armar el producto final."
                ),
                "sync_url": reverse("recetas:receta_sync_derivados", args=[source_base_receta.id]),
                "detail_url": reverse("recetas:receta_detail", args=[source_base_receta.id]),
            }

    if request.method == "POST":
        values = {
            "nombre": (request.POST.get("nombre") or "").strip(),
            "codigo_point": (request.POST.get("codigo_point") or "").strip(),
            "familia": (request.POST.get("familia") or "").strip(),
            "categoria": (request.POST.get("categoria") or "").strip(),
            "sheet_name": (request.POST.get("sheet_name") or "").strip(),
            "recipe_mode": (request.POST.get("recipe_mode") or "BASE").strip().upper(),
            "rendimiento_cantidad": (request.POST.get("rendimiento_cantidad") or "").strip(),
            "rendimiento_unidad_id": (request.POST.get("rendimiento_unidad_id") or "").strip(),
        }
        mode = values["recipe_mode"] if values["recipe_mode"] in {"BASE", "BASE_DERIVADOS", "FINAL"} else "BASE"

        if not values["nombre"]:
            messages.error(request, "El nombre de la receta es obligatorio.")
        else:
            validation_errors = _validate_receta_enterprise_rules(
                mode_code=mode,
                usa_presentaciones=(mode == "BASE_DERIVADOS"),
                rendimiento_cantidad=_to_decimal_or_none(values["rendimiento_cantidad"]),
                rendimiento_unidad_id=values["rendimiento_unidad_id"],
                familia=values["familia"],
            )
            if validation_errors:
                for error in validation_errors:
                    messages.error(request, error)
            else:
                rendimiento_cantidad = (
                    None if mode == "FINAL" else _to_decimal_or_none(values["rendimiento_cantidad"])
                )
                rendimiento_unidad = (
                    None
                    if mode == "FINAL" or not values["rendimiento_unidad_id"]
                    else UnidadMedida.objects.filter(pk=values["rendimiento_unidad_id"]).first()
                )
                receta = Receta.objects.create(
                    nombre=values["nombre"][:250],
                    codigo_point=values["codigo_point"][:80],
                    familia=values["familia"][:120],
                    categoria=values["categoria"][:120],
                    sheet_name=values["sheet_name"][:120],
                    tipo=Receta.TIPO_PRODUCTO_FINAL if mode == "FINAL" else Receta.TIPO_PREPARACION,
                    usa_presentaciones=(mode == "BASE_DERIVADOS"),
                    rendimiento_cantidad=rendimiento_cantidad,
                    rendimiento_unidad=rendimiento_unidad,
                    hash_contenido=uuid4().hex,
                )
                log_event(
                    request.user,
                    "CREATE_RECETA",
                    "recetas.Receta",
                    str(receta.pk),
                    {
                        "nombre": receta.nombre,
                        "tipo": receta.tipo,
                        "usa_presentaciones": receta.usa_presentaciones,
                        "familia": receta.familia,
                        "categoria": receta.categoria,
                    },
                )
                if mode == "FINAL":
                    messages.success(
                        request,
                        "Producto final creado. Ahora agrega sus componentes de armado.",
                    )
                elif mode == "BASE_DERIVADOS":
                    messages.success(
                        request,
                        "Insumo base creado. El siguiente paso es definir sus presentaciones derivadas.",
                    )
                else:
                    messages.success(
                        request,
                        "Insumo interno base creado. Ahora captura su receta y rendimiento.",
                    )
                return redirect("recetas:receta_detail", pk=receta.pk)

        defaults.update(values)

    return render(
        request,
        "recetas/receta_create.html",
        {
            "recipe_modes": recipe_modes,
            "familias_catalogo": familias_catalogo,
            "categorias_catalogo": categorias_catalogo,
            "familia_categoria_catalogo_json": familia_categoria_catalogo_json(categorias_catalogo),
            "unidades": unidades,
            "values": defaults,
            "source_base_receta": source_base_receta,
            "source_base_context": source_base_context,
        },
    )


@login_required
def receta_detail(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    familias_db = list(
        Receta.objects.exclude(familia__exact="")
        .values_list("familia", flat=True)
        .distinct()
        .order_by("familia")
    )
    familias_catalogo = familias_producto_catalogo(familias_db)
    categorias_catalogo = list(
        Receta.objects.exclude(categoria__exact="")
        .values_list("categoria", flat=True)
        .distinct()
        .order_by("categoria")
    )
    versiones_unavailable = False
    versiones_all = []
    versiones_recientes = []
    comparativo = None
    try:
        if not receta.versiones_costo.exists():
            _sync_cost_version_safe(request, receta, "AUTO_BOOTSTRAP")
        versiones_all = _load_versiones_costeo(receta, 80)
        versiones_recientes = versiones_all[:8]
        comparativo = comparativo_versiones(versiones_recientes)
    except (OperationalError, ProgrammingError):
        versiones_unavailable = True
        messages.warning(
            request,
            "Histórico de versiones de costo no disponible en este entorno. Ejecuta migraciones para habilitarlo.",
        )

    lineas = list(receta.lineas.select_related("insumo").order_by("posicion"))
    presentaciones = sorted(list(receta.presentaciones.all()), key=lambda p: _presentacion_sort_key(p.nombre))
    costeo_unavailable = False
    try:
        costeo_actual = calcular_costeo_receta(receta)
    except (OperationalError, ProgrammingError):
        costeo_unavailable = True
        costeo_actual = _empty_costeo_actual()
        messages.warning(
            request,
            "Costeo avanzado no disponible en este entorno. Ejecuta migraciones para habilitar drivers.",
        )

    selected_base = None
    selected_target = None
    compare_data = None
    base_raw = (request.GET.get("base") or "").strip()
    target_raw = (request.GET.get("target") or "").strip()
    try:
        if base_raw and target_raw:
            base_num = int(base_raw)
            target_num = int(target_raw)
            by_num = {v.version_num: v for v in versiones_all}
            if base_num in by_num and target_num in by_num and base_num != target_num:
                selected_base = by_num[base_num]
                selected_target = by_num[target_num]
    except Exception:
        selected_base = None
        selected_target = None

    if not selected_base and not selected_target and len(versiones_all) >= 2:
        selected_target = versiones_all[0]
        selected_base = versiones_all[1]

    if selected_base and selected_target:
        compare_data = _compare_versions(selected_base, selected_target)

    def _line_origin_bucket(linea: LineaReceta) -> str:
        if linea.tipo_linea == LineaReceta.TIPO_SUBSECCION:
            return "subsecciones"
        if not linea.insumo:
            return "pendientes"
        if linea.insumo.tipo_item == Insumo.TIPO_EMPAQUE:
            return "empaques"
        if linea.insumo.tipo_item == Insumo.TIPO_INTERNO or (linea.insumo.codigo or "").startswith("DERIVADO:RECETA:"):
            return "internos"
        return "materia_prima"

    total_lineas = len(lineas)
    total_match = sum(1 for l in lineas if l.match_status == LineaReceta.STATUS_AUTO)
    total_revision = sum(1 for l in lineas if l.match_status == LineaReceta.STATUS_NEEDS_REVIEW)
    total_sin_match = sum(1 for l in lineas if l.match_status == LineaReceta.STATUS_REJECTED)
    total_costo_directo = sum((l.costo_total_estimado or 0.0) for l in lineas)
    total_costo_estimado = float(receta.costo_total_estimado or 0.0)
    total_materia_prima = sum(1 for l in lineas if l.insumo and l.insumo.tipo_item == Insumo.TIPO_MATERIA_PRIMA)
    total_internos = sum(
        1
        for l in lineas
        if l.insumo and (l.insumo.tipo_item == Insumo.TIPO_INTERNO or (l.insumo.codigo or "").startswith("DERIVADO:RECETA:"))
    )
    total_empaques = sum(1 for l in lineas if l.insumo and l.insumo.tipo_item == Insumo.TIPO_EMPAQUE)
    total_subsecciones = sum(1 for l in lineas if l.tipo_linea == LineaReceta.TIPO_SUBSECCION)
    receta_rol_label = (
        "Producto final de venta"
        if receta.tipo == Receta.TIPO_PRODUCTO_FINAL
        else "Insumo interno / batida base"
    )
    receta_rol_description = (
        "Se arma con subinsumos, insumos internos, materia prima puntual y empaques."
        if receta.tipo == Receta.TIPO_PRODUCTO_FINAL
        else "Se produce con materia prima y puede generar subinsumos o presentaciones derivadas."
    )
    is_producto_final = receta.tipo == Receta.TIPO_PRODUCTO_FINAL
    is_base_con_presentaciones = receta.tipo == Receta.TIPO_PREPARACION and receta.usa_presentaciones
    is_base_simple = receta.tipo == Receta.TIPO_PREPARACION and not receta.usa_presentaciones
    supply_chain_snapshot = _recipe_supply_chain_snapshot(receta)
    active_presentaciones = [p for p in presentaciones if p.activo]
    presentacion_derivados_activos = Insumo.objects.filter(
        codigo__startswith=f"DERIVADO:RECETA:{receta.id}:PRESENTACION:",
        activo=True,
    ).count()
    add_line_label = "Agregar componente" if is_producto_final else "Agregar insumo base"
    copy_label = "Copiar estructura base" if is_producto_final else "Copiar ingredientes"
    chain_focus_summary = _recipe_chain_focus_summary(receta)
    enterprise_stage = _recipe_enterprise_stage(receta)
    enterprise_stage_playbook = _recipe_enterprise_stage_playbook(receta)
    next_steps = []
    chain_actions: list[dict[str, str]] = []
    base_chain_actions: list[dict[str, str]] = []
    base_chain_checkpoints: list[dict[str, object]] = []
    if is_base_simple:
        next_steps = [
            "Captura materias primas y cantidades reales de la batida.",
            "Define rendimiento para calcular costo por unidad base.",
            "Si después saldrán tamaños, activa presentaciones y crea los derivados.",
        ]
        base_chain_checkpoints = [
            {
                "label": "Rendimiento capturado",
                "ok": bool(receta.rendimiento_cantidad and receta.rendimiento_cantidad > 0),
                "detail": "Costo unitario base listo para operar." if receta.rendimiento_cantidad else "Todavía falta rendimiento total.",
            },
            {
                "label": "Unidad de rendimiento",
                "ok": bool(receta.rendimiento_unidad_id),
                "detail": receta.rendimiento_unidad.codigo if receta.rendimiento_unidad_id else "Sin unidad base definida.",
            },
        ]
    elif is_base_con_presentaciones:
        next_steps = [
            "Captura la receta completa de la batida o mezcla base.",
            "Mantén actualizado el rendimiento total para costeo por kg/lt/pza.",
            "Administra presentaciones derivadas para que aparezcan como subinsumos en productos finales.",
        ]
        prep_insumo_operativo = bool(supply_chain_snapshot and supply_chain_snapshot["prep_insumo"])
        chain_actions.append(
            {"label": "Agregar presentación", "url": reverse("recetas:presentacion_create", args=[receta.id])}
        )
        if active_presentaciones and presentacion_derivados_activos < len(active_presentaciones):
            chain_actions.append(
                {"label": "Sincronizar derivados", "url": reverse("recetas:receta_sync_derivados", args=[receta.id])}
            )
        if supply_chain_snapshot and not supply_chain_snapshot["has_downstream_usage"]:
            chain_actions.append(
                {
                    "label": "Crear producto final",
                    "url": f"{reverse('recetas:receta_create')}?mode=FINAL&source_base={receta.id}",
                }
            )
        sync_missing_count = max(len(active_presentaciones) - presentacion_derivados_activos, 0)
        base_chain_actions.append(
            {"label": "Administrar presentaciones", "url": reverse("recetas:presentacion_create", args=[receta.id])}
        )
        if not prep_insumo_operativo or sync_missing_count > 0:
            base_chain_actions.append(
                {
                    "label": "Sincronizar derivados",
                    "url": f"{reverse('recetas:receta_sync_derivados', args=[receta.id])}?next={reverse('recetas:receta_detail', args=[receta.id])}",
                }
            )
        if supply_chain_snapshot and not supply_chain_snapshot["has_downstream_usage"] and active_presentaciones:
            base_chain_actions.append(
                {
                    "label": "Crear producto final",
                    "url": f"{reverse('recetas:receta_create')}?mode=FINAL&source_base={receta.id}",
                }
            )
        base_chain_checkpoints = [
            {
                "label": "Insumo base derivado",
                "ok": prep_insumo_operativo,
                "detail": "Ya existe el insumo base maestro." if prep_insumo_operativo else "Falta generar o reactivar el derivado base.",
            },
            {
                "label": "Presentaciones activas",
                "ok": len(active_presentaciones) > 0,
                "detail": f"{len(active_presentaciones)} activa(s)." if active_presentaciones else "Todavía no hay tamaños activos.",
            },
            {
                "label": "Derivados activos",
                "ok": sync_missing_count == 0 and len(active_presentaciones) > 0,
                "detail": (
                    f"{presentacion_derivados_activos}/{len(active_presentaciones)} derivados listos."
                    if active_presentaciones
                    else "No aplica hasta crear presentaciones."
                ),
            },
            {
                "label": "Consumo final",
                "ok": bool(supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"]),
                "detail": (
                    f"{supply_chain_snapshot['downstream_recipe_count']} producto(s) final(es) usan esta base."
                    if supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"]
                    else "Aún no hay producto final consumiendo esta base."
                ),
            },
        ]
    else:
        next_steps = [
            "Agrega solo componentes maestros: panes, rellenos, coberturas, empaques o MP puntual.",
            "Usa copiar estructura si este producto se parece a otro ya armado.",
            "Ajusta cantidades finales; el sistema toma unidad y costo desde el maestro de artículos.",
        ]
    recetas_copiables = (
        Receta.objects.exclude(pk=receta.pk)
        .order_by("tipo", "nombre")
        .only("id", "nombre", "tipo", "familia", "categoria")
    )
    ordered_group_keys = (
        ["internos", "materia_prima", "empaques", "subsecciones", "pendientes"]
        if is_producto_final
        else ["materia_prima", "internos", "empaques", "subsecciones", "pendientes"]
    )
    group_meta = {
        "internos": {
            "title": "Insumos internos / subinsumos",
            "description": "Componentes producidos dentro de la empresa y reutilizados en esta receta.",
            "badge": "Interno",
        },
        "materia_prima": {
            "title": "Materia prima directa",
            "description": "Componentes comprados a proveedor y consumidos directamente en la receta.",
            "badge": "Materia prima",
        },
        "empaques": {
            "title": "Empaques",
            "description": "Material de presentación o empaque final ligado al costo del producto.",
            "badge": "Empaque",
        },
        "subsecciones": {
            "title": "Subsecciones / bloques",
            "description": "Bloques operativos de armado para ordenar coberturas, rellenos y decoración.",
            "badge": "Subsección",
        },
        "pendientes": {
            "title": "Sin artículo estándar",
            "description": "Componentes todavía sin artículo estándar asignado. Deben validarse para costeo correcto.",
            "badge": "Por validar",
        },
    }
    grouped_line_map: dict[str, list[LineaReceta]] = {key: [] for key in ordered_group_keys}
    direct_base_replacement_cache: dict[int, list[dict[str, object]]] = {}
    source_recipe_ids = {
        recipe_id
        for linea in lineas
        if linea.insumo_id
        for recipe_id in [_recipe_id_from_derived_code(linea.insumo.codigo or "")]
        if recipe_id
    }
    source_recipe_map = {
        item.id: item
        for item in Receta.objects.filter(id__in=source_recipe_ids).only("id", "nombre", "usa_presentaciones")
    }
    source_recipe_presentaciones = {
        row["receta_id"]: row["total"]
        for row in (
            RecetaPresentacion.objects.filter(receta_id__in=source_recipe_ids, activo=True)
            .values("receta_id")
            .annotate(total=Count("id"))
        )
    }
    for linea in lineas:
        _attach_linea_suggested_match(linea)
        _attach_linea_canonical_target(linea)
        linea.source_recipe = None
        linea.source_code_kind = None
        linea.source_active_presentaciones_count = 0
        linea.uses_direct_base_in_final = False
        linea.direct_base_replacement = None
        linea.erp_profile = _insumo_erp_readiness(linea.insumo) if linea.insumo_id else None
        if linea.insumo_id:
            linea.source_code_kind = _derived_code_kind(linea.insumo.codigo or "")
            source_recipe_id = _recipe_id_from_derived_code(linea.insumo.codigo or "")
            if source_recipe_id:
                linea.source_recipe = source_recipe_map.get(source_recipe_id)
                linea.source_active_presentaciones_count = int(
                    source_recipe_presentaciones.get(source_recipe_id, 0)
                )
                linea.uses_direct_base_in_final = bool(
                    is_producto_final
                    and linea.insumo.tipo_item == Insumo.TIPO_INTERNO
                    and linea.source_code_kind == "PREPARACION"
                    and linea.source_recipe
                    and linea.source_recipe.usa_presentaciones
                    and linea.source_active_presentaciones_count > 0
                )
                if linea.uses_direct_base_in_final:
                    linea.direct_base_replacement = _suggest_direct_base_replacement(
                        linea,
                        cache=direct_base_replacement_cache,
                    )
        grouped_line_map[_line_origin_bucket(linea)].append(linea)
    line_groups = [
        {
            "key": key,
            "title": group_meta[key]["title"],
            "description": group_meta[key]["description"],
            "badge": group_meta[key]["badge"],
            "lineas": grouped_line_map[key],
            "count": len(grouped_line_map[key]),
            "subtotal": sum((l.costo_total_estimado or 0.0) for l in grouped_line_map[key]),
            "erp_ready_count": sum(
                1
                for l in grouped_line_map[key]
                if getattr(l, "erp_profile", None) and l.erp_profile["ready"]
            ),
            "erp_incomplete_count": sum(
                1
                for l in grouped_line_map[key]
                if getattr(l, "erp_profile", None) and not l.erp_profile["ready"]
            ),
        }
        for key in ordered_group_keys
        if grouped_line_map[key]
    ]
    recipe_master_blockers = _recipe_master_blockers(lineas, receta=receta)
    recipe_master_gap_totals = _recipe_master_gap_totals_from_blockers(recipe_master_blockers)
    component_breakdown = []
    product_upstream_snapshot = _product_upstream_snapshot(lineas, receta=receta) if is_producto_final else None
    if product_upstream_snapshot is not None:
        setattr(receta, "_product_upstream_snapshot_cache", product_upstream_snapshot)
    derived_parent_snapshot = (
        product_upstream_snapshot.get("derived_parent_snapshot")
        if product_upstream_snapshot is not None
        else None
    )
    non_canonical_count = sum(1 for l in lineas if getattr(l, "canonical_needs_repoint", False))
    if total_costo_estimado > 0:
        if derived_parent_snapshot:
            derived_parent_cost = float(derived_parent_snapshot.get("parent_unit_cost") or 0.0)
            if derived_parent_cost > 0:
                component_breakdown.append(
                    {
                        "key": "parent_base_prorrated",
                        "title": "Base padre prorrateada",
                        "accent": "is-primary",
                        "subtotal": derived_parent_cost,
                        "percentage": (derived_parent_cost / total_costo_estimado * 100.0),
                        "count": 1,
                        "erp_ready_count": 1,
                        "erp_incomplete_count": 0,
                    }
                )
        breakdown_meta = {
            "internos": {"title": "Insumos internos", "accent": "is-success"},
            "materia_prima": {"title": "Materia prima puntual", "accent": "is-warning"},
            "empaques": {"title": "Empaques", "accent": "is-info"},
        }
        for key in ("internos", "materia_prima", "empaques"):
            subtotal = sum((l.costo_total_estimado or 0.0) for l in grouped_line_map[key])
            if subtotal <= 0:
                continue
            percentage = (subtotal / total_costo_estimado * 100.0) if total_costo_estimado else 0.0
            component_breakdown.append(
                {
                    "key": key,
                    "title": breakdown_meta[key]["title"],
                    "accent": breakdown_meta[key]["accent"],
                    "subtotal": subtotal,
                    "percentage": percentage,
                    "count": len(grouped_line_map[key]),
                    "erp_ready_count": sum(
                        1
                        for l in grouped_line_map[key]
                        if getattr(l, "erp_profile", None) and l.erp_profile["ready"]
                    ),
                    "erp_incomplete_count": sum(
                        1
                        for l in grouped_line_map[key]
                        if getattr(l, "erp_profile", None) and not l.erp_profile["ready"]
                    ),
                }
            )
    bom_integrity_alerts = []
    direct_base_usage_lines = [linea for linea in lineas if getattr(linea, "uses_direct_base_in_final", False)]
    direct_base_suggested_count = sum(
        1 for linea in direct_base_usage_lines if getattr(linea, "direct_base_replacement", None)
    )
    incomplete_item_count = sum(
        1 for l in lineas if getattr(l, "erp_profile", None) and not l.erp_profile["ready"]
    )
    prep_code = f"DERIVADO:RECETA:{receta.id}:PREPARACION"
    presentation_code_prefix = f"DERIVADO:RECETA:{receta.id}:PRESENTACION:"
    active_presentaciones = [p for p in presentaciones if p.activo]
    prep_insumo_ready = Insumo.objects.filter(codigo=prep_code, activo=True).exists()
    presentacion_derivados_activos = Insumo.objects.filter(
        codigo__startswith=presentation_code_prefix,
        activo=True,
    ).count()
    presentacion_health = None
    if receta.tipo == Receta.TIPO_PREPARACION:
        has_rendimiento = bool(receta.rendimiento_cantidad and receta.rendimiento_cantidad > 0)
        has_rendimiento_unidad = bool(receta.rendimiento_unidad_id)
        readiness_label = "Listo para derivados"
        readiness_level = "success"
        readiness_note = "La base ya tiene rendimiento, unidad y derivados suficientes para alimentar productos finales."
        if not has_rendimiento or not has_rendimiento_unidad:
            readiness_label = "Falta rendimiento base"
            readiness_level = "danger"
            readiness_note = "Captura rendimiento y unidad para calcular costo unitario y habilitar derivados consistentes."
        elif receta.usa_presentaciones and not active_presentaciones:
            readiness_label = "Sin presentaciones activas"
            readiness_level = "warning"
            readiness_note = "La receta está marcada con derivados, pero todavía no hay tamaños activos listos para usarse."
        elif receta.usa_presentaciones and presentacion_derivados_activos < len(active_presentaciones):
            readiness_label = "Derivados por sincronizar"
            readiness_level = "warning"
            readiness_note = "Hay presentaciones activas sin todos sus insumos derivados activos. Conviene revisar la sincronización."

        presentacion_health = {
            "readiness_label": readiness_label,
            "readiness_level": readiness_level,
            "readiness_note": readiness_note,
            "has_rendimiento": has_rendimiento,
            "has_rendimiento_unidad": has_rendimiento_unidad,
            "prep_insumo_ready": prep_insumo_ready,
            "active_presentaciones_count": len(active_presentaciones),
            "derived_presentaciones_count": presentacion_derivados_activos,
            "sync_recommended": bool(
                receta.usa_presentaciones
                and active_presentaciones
                and (not prep_insumo_ready or presentacion_derivados_activos < len(active_presentaciones))
            ),
        }
    if is_producto_final:
        if total_internos == 0:
            bom_integrity_alerts.append(
                {
                    "level": "danger",
                    "title": "Faltan insumos internos",
                    "description": "El producto final no tiene panes, rellenos, coberturas o subinsumos ligados. Esto suele indicar un BOM incompleto.",
                    "action_label": "Agregar insumo interno",
                    "action_url": f"{reverse('recetas:linea_create', args=[receta.id])}?component_kind=INSUMO_INTERNO&component_context=internos",
                }
            )
        if total_empaques == 0:
            bom_integrity_alerts.append(
                {
                    "level": "warning",
                    "title": "Sin empaque ligado",
                    "description": "No hay domos, cajas, etiquetas ni material final en el costeo. Revisa si el producto debe incluir empaque.",
                    "action_label": "Agregar empaque",
                    "action_url": f"{reverse('recetas:linea_create', args=[receta.id])}?component_kind=EMPAQUE&component_context=empaques",
                }
            )
        if total_lineas > 0 and total_internos == 0 and total_materia_prima > 0:
            bom_integrity_alerts.append(
                {
                    "level": "warning",
                    "title": "Producto armado directo con materia prima",
                    "description": "Este producto final está consumiendo solo materia prima puntual. Confirma que no deba usar insumos internos ya estandarizados.",
                }
            )
        if product_upstream_snapshot and product_upstream_snapshot["internal_without_source_count"] > 0:
            bom_integrity_alerts.append(
                {
                    "level": "warning",
                    "title": "Insumos internos sin base origen",
                    "description": f"Se detectaron {product_upstream_snapshot['internal_without_source_count']} línea(s) internas sin trazabilidad a receta base. Conviene normalizarlas para estabilizar costeo, MRP y compras.",
                }
            )
        if direct_base_usage_lines:
            ejemplos = ", ".join(
                dict.fromkeys(
                    (
                        getattr(linea.source_recipe, "nombre", "")
                        or getattr(linea.insumo, "nombre", "")
                    )
                    for linea in direct_base_usage_lines[:3]
                )
            )
            description = (
                f"Se detectaron {len(direct_base_usage_lines)} componente(s) usando la base completa "
                "aunque la receta origen ya maneja presentaciones activas. Sustituye esa base por la "
                "presentación derivada correcta para estabilizar costo y MRP."
            )
            if ejemplos:
                description += f" Bases afectadas: {ejemplos}."
            bom_integrity_alerts.append(
                {
                    "level": "danger",
                    "title": "Producto final usando base sin presentación",
                    "description": description,
                    "action_label": "Revisar componentes",
                    "action_url": reverse("recetas:receta_detail", args=[receta.id]),
                }
            )
        if total_revision or total_sin_match:
            bom_integrity_alerts.append(
                {
                    "level": "danger",
                    "title": "Hay componentes sin artículo estándar",
                    "description": "Mientras existan componentes por validar o sin artículo estándar, el costo y la planeación pueden quedar inestables.",
                    "action_label": "Resolver catálogo",
                    "action_url": f"{reverse('recetas:matching_pendientes')}?receta={receta.id}",
                }
            )
        if non_canonical_count:
            bom_integrity_alerts.append(
                {
                    "level": "warning",
                    "title": "Hay artículos fuera de maestro",
                    "description": f"Se detectaron {non_canonical_count} línea(s) apuntando a variantes fuera de maestro. Conviene corregirlas para estabilizar costos, compras y reportes.",
                    "action_label": "Usar artículo estándar",
                    "action_url": reverse("recetas:receta_detail", args=[receta.id]),
                }
            )
        if incomplete_item_count:
            bom_integrity_alerts.append(
                {
                    "level": "danger" if recipe_master_gap_totals["critical"] > 0 else "warning",
                    "title": (
                        "Demanda crítica bloqueada por maestro"
                        if recipe_master_gap_totals["critical"] > 0
                        else "Hay artículos incompletos en maestro"
                    ),
                    "description": (
                        f"Se detectaron {recipe_master_gap_totals['critical']} artículo(s) críticos por demanda con brechas en maestro. Conviene corregirlos antes de cerrar costeo, MRP o compras."
                        if recipe_master_gap_totals["critical"] > 0
                        else f"Se detectaron {incomplete_item_count} componente(s) ligados a artículos todavía incompletos en el maestro. Conviene corregirlos antes de cerrar costeo o MRP."
                    ),
                    "action_label": "Revisar maestro",
                    "action_url": f"{reverse('maestros:insumo_list')}?enterprise_status=incompletos&usage_scope=recipes&recipe_scope=finales",
                }
            )
    elif is_base_con_presentaciones:
        if not receta.rendimiento_cantidad:
            bom_integrity_alerts.append(
                {
                    "level": "danger",
                    "title": "Falta rendimiento base",
                    "description": "Sin rendimiento no puede calcularse el costo por unidad ni el costo de las presentaciones derivadas.",
                    "action_label": "Abrir receta",
                    "action_url": reverse("recetas:receta_detail", args=[receta.id]),
                }
            )
        if not presentaciones:
            bom_integrity_alerts.append(
                {
                    "level": "warning",
                    "title": "Sin presentaciones activas",
                    "description": "La receta está marcada para derivados, pero aún no tiene presentaciones configuradas.",
                    "action_label": "Agregar presentación",
                    "action_url": reverse("recetas:presentacion_create", args=[receta.id]),
                }
            )
        elif presentacion_derivados_activos < len(active_presentaciones):
            bom_integrity_alerts.append(
                {
                    "level": "warning",
                    "title": "Derivados pendientes de sincronizar",
                    "description": "Hay presentaciones activas que todavía no generaron todos sus insumos derivados maestros. Conviene sincronizar antes de usarlas en productos finales.",
                    "action_label": "Sincronizar derivados",
                    "action_url": reverse("recetas:receta_sync_derivados", args=[receta.id]),
                }
            )
        elif not _recipe_supply_chain_snapshot(receta)["has_downstream_usage"]:
            bom_integrity_alerts.append(
                {
                    "level": "warning",
                    "title": "Sin consumo final",
                    "description": "La base ya tiene derivados, pero todavía no se consume en ningún producto final. Conviene cerrar la cadena operativa.",
                    "action_label": "Crear producto final",
                    "action_url": f"{reverse('recetas:receta_create')}?mode=FINAL&source_base={receta.id}",
                }
            )
    elif is_base_simple and not receta.rendimiento_cantidad:
        bom_integrity_alerts.append(
            {
                "level": "warning",
                "title": "Falta rendimiento",
                "description": "Captura el rendimiento total para que el costo unitario de la batida quede bien calculado.",
                "action_label": "Abrir receta",
                "action_url": reverse("recetas:receta_detail", args=[receta.id]),
            }
        )
    recipe_chain_status = _recipe_chain_focus_summary(receta)
    governance_issues = _recipe_governance_issues(receta)
    release_gate_rows = [
        {
            "label": "Maestro ERP cerrado",
            "status": recipe_master_gap_totals["total"] == 0,
            "detail": (
                "Todos los artículos ligados ya tienen la base maestra completa."
                if recipe_master_gap_totals["total"] == 0
                else f'{recipe_master_gap_totals["total"]} artículo(s) aún requieren datos maestros.'
            ),
            "action_label": "Abrir maestro",
            "action_url": reverse("maestros:insumo_list"),
        },
        {
            "label": "BOM sin componentes abiertos",
            "status": total_revision == 0 and total_sin_match == 0,
            "detail": (
                "La estructura ya quedó integrada y sin componentes abiertos."
                if total_revision == 0 and total_sin_match == 0
                else f"{total_revision} componente(s) por validar · {total_sin_match} sin artículo estándar."
            ),
            "action_label": add_line_label,
            "action_url": reverse("recetas:linea_create", args=[receta.id]),
        },
        {
            "label": "Costo operativo calculado",
            "status": total_costo_estimado > 0,
            "detail": (
                f"Costo operativo actual ${total_costo_estimado:.2f}."
                if total_costo_estimado > 0
                else "Todavía no hay costo operativo suficiente para cerrar el documento."
            ),
            "action_label": "Abrir costeo",
            "action_url": reverse("recetas:drivers_costeo"),
        },
        {
            "label": "Cadena ERP cerrada",
            "status": recipe_chain_status["tone"] != "danger",
            "detail": recipe_chain_status["summary"],
            "action_label": chain_actions[0]["label"] if chain_actions else "Revisar documento",
            "action_url": chain_actions[0]["url"] if chain_actions else reverse("recetas:receta_detail", args=[receta.id]),
        },
    ]
    release_gate_progress = {
        "completed": sum(1 for row in release_gate_rows if row["status"]),
        "total": len(release_gate_rows),
    }
    erp_governance_rows = _recipe_detail_governance_rows(
        receta=receta,
        recipe_master_gap_totals=recipe_master_gap_totals,
        total_revision=total_revision,
        total_sin_match=total_sin_match,
        release_gate_progress=release_gate_progress,
        chain_focus_summary=chain_focus_summary,
    )
    trunk_handoff_rows = _recipe_detail_trunk_handoff_rows(
        receta=receta,
        recipe_master_gap_totals=recipe_master_gap_totals,
        total_revision=total_revision,
        total_sin_match=total_sin_match,
        release_gate_progress=release_gate_progress,
        chain_focus_summary=chain_focus_summary,
        is_producto_final=is_producto_final,
        is_base_con_presentaciones=is_base_con_presentaciones,
        governance_issues=governance_issues,
        supply_chain_snapshot=supply_chain_snapshot,
        product_upstream_snapshot=product_upstream_snapshot,
    )
    erp_control_chain = [
        {
            "label": "01 Maestro",
            "status": "Listo" if recipe_master_gap_totals["total"] == 0 else "Con brechas",
            "tone": "success" if recipe_master_gap_totals["total"] == 0 else "warning",
            "detail": (
                "Los artículos ligados ya cumplen datos base."
                if recipe_master_gap_totals["total"] == 0
                else f'{recipe_master_gap_totals["total"]} artículo(s) requieren cierre maestro.'
            ),
            "action_label": "Abrir maestro",
            "action_url": reverse("maestros:insumo_list"),
        },
        {
            "label": "02 BOM",
            "status": "Listo" if total_revision == 0 and total_sin_match == 0 else "Por validar",
            "tone": "success" if total_revision == 0 and total_sin_match == 0 else "warning",
            "detail": (
                f"{total_lineas} componente(s) ya estructurados."
                if total_revision == 0 and total_sin_match == 0
                else f"{total_revision} componente(s) por validar · {total_sin_match} sin artículo estándar."
            ),
            "action_label": add_line_label,
            "action_url": reverse("recetas:linea_create", args=[receta.id]),
        },
        {
            "label": "03 Costeo",
            "status": "Calculado" if total_costo_estimado > 0 else "Por validar",
            "tone": "success" if total_costo_estimado > 0 else "warning",
            "detail": (
                f"Costo operativo actual ${total_costo_estimado:.2f}."
                if total_costo_estimado > 0
                else "Todavía no hay costo operativo suficiente para cerrar el documento."
            ),
            "action_label": "Abrir costeo",
            "action_url": reverse("recetas:drivers_costeo"),
        },
        {
            "label": "04 Liberación",
            "status": (
                "Lista"
                if recipe_master_gap_totals["total"] == 0
                and total_revision == 0
                and total_sin_match == 0
                and total_costo_estimado > 0
                and recipe_chain_status["tone"] != "danger"
                else "Por validar"
            ),
            "tone": (
                "success"
                if recipe_master_gap_totals["total"] == 0
                and total_revision == 0
                and total_sin_match == 0
                and total_costo_estimado > 0
                and recipe_chain_status["tone"] != "danger"
                else "warning"
            ),
            "detail": (
                "Documento listo para costeo, planeación y operación diaria."
                if recipe_master_gap_totals["total"] == 0
                and total_revision == 0
                and total_sin_match == 0
                and total_costo_estimado > 0
                and recipe_chain_status["tone"] != "danger"
                else recipe_chain_status["summary"]
            ),
            "action_label": chain_actions[0]["label"] if chain_actions else "Revisar documento",
            "action_url": chain_actions[0]["url"] if chain_actions else reverse("recetas:receta_detail", args=[receta.id]),
        },
    ]
    module_enablement_cards: list[dict[str, str | bool]] = []
    master_ready = recipe_master_gap_totals["total"] == 0
    bom_ready = total_revision == 0 and total_sin_match == 0
    cost_ready = total_costo_estimado > 0
    chain_ready = recipe_chain_status["tone"] != "danger"
    packaging_ready = total_empaques > 0
    internal_ready = total_internos > 0
    if is_producto_final:
        module_enablement_cards = [
            {
                "label": "Costeo",
                "status": master_ready and bom_ready and cost_ready,
                "tone": "success" if master_ready and bom_ready and cost_ready else "warning",
                "detail": (
                    "El documento ya tiene base maestra, BOM estable y costo repetible."
                    if master_ready and bom_ready and cost_ready
                    else "Todavía falta cerrar maestro, estructura o costo para usar este producto como referencia financiera."
                ),
                "action_label": "Abrir costeo" if not cost_ready else "Revisar documento",
                "action_url": reverse("recetas:drivers_costeo") if not cost_ready else reverse("recetas:receta_detail", args=[receta.id]),
            },
            {
                "label": "MRP",
                "status": master_ready and bom_ready and cost_ready and chain_ready and internal_ready,
                "tone": "success" if master_ready and bom_ready and cost_ready and chain_ready and internal_ready else "warning",
                "detail": (
                    "El producto ya puede explotar demanda y reabasto con trazabilidad cerrada."
                    if master_ready and bom_ready and cost_ready and chain_ready and internal_ready
                    else "MRP sigue bloqueado hasta cerrar componentes internos, maestro y trazabilidad del documento."
                ),
                "action_label": "Abrir MRP" if master_ready and bom_ready and cost_ready and chain_ready and internal_ready else "Cerrar documento",
                "action_url": reverse("recetas:mrp_form") if master_ready and bom_ready and cost_ready and chain_ready and internal_ready else reverse("recetas:receta_detail", args=[receta.id]),
            },
            {
                "label": "Compras",
                "status": master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready,
                "tone": "success" if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready else "warning",
                "detail": (
                    "La receta ya puede alimentar solicitudes y órdenes sin ambigüedad documental."
                    if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready
                    else "Compras sigue bloqueado hasta cerrar empaque, estructura BOM y artículos del maestro."
                ),
                "action_label": "Abrir compras" if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready else "Revisar bloqueos",
                "action_url": reverse("compras:solicitudes") if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready else reverse("recetas:receta_detail", args=[receta.id]),
            },
            {
                "label": "Inventario",
                "status": master_ready and chain_ready and packaging_ready,
                "tone": "success" if master_ready and chain_ready and packaging_ready else "warning",
                "detail": (
                    "Inventario ya puede trabajar con el producto y su empaque como artículo operativo estable."
                    if master_ready and chain_ready and packaging_ready
                    else "Inventario aún requiere maestro completo, trazabilidad cerrada o empaque definido."
                ),
                "action_label": "Abrir inventario" if master_ready and chain_ready and packaging_ready else "Revisar maestro",
                "action_url": reverse("inventario:existencias") if master_ready and chain_ready and packaging_ready else reverse("maestros:insumo_list"),
            },
            {
                "label": "Venta final",
                "status": master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready,
                "tone": "success" if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready else "warning",
                "detail": (
                    "El artículo ya está en condición de venta final consistente."
                    if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready
                    else "Todavía no conviene liberar este artículo a operación final hasta cerrar estructura, costo y empaque."
                ),
                "action_label": "Revisar catálogo" if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready else "Cerrar ERP",
                "action_url": reverse("recetas:recetas_list",) if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready else reverse("recetas:receta_detail", args=[receta.id]),
            },
        ]
    elif is_base_con_presentaciones:
        module_enablement_cards = [
            {
                "label": "Costeo base",
                "status": bool(presentacion_health and presentacion_health["has_rendimiento"] and presentacion_health["has_rendimiento_unidad"]),
                "tone": "success" if presentacion_health and presentacion_health["has_rendimiento"] and presentacion_health["has_rendimiento_unidad"] else "warning",
                "detail": (
                    "La base ya tiene rendimiento y unidad para costeo repetible."
                    if presentacion_health and presentacion_health["has_rendimiento"] and presentacion_health["has_rendimiento_unidad"]
                    else "Todavía falta rendimiento o unidad para que la base costee correctamente."
                ),
                "action_label": "Abrir base",
                "action_url": reverse("recetas:receta_detail", args=[receta.id]),
            },
            {
                "label": "Derivados",
                "status": bool(presentacion_health and presentacion_health["active_presentaciones_count"] > 0 and presentacion_health["derived_presentaciones_count"] >= presentacion_health["active_presentaciones_count"]),
                "tone": "success" if presentacion_health and presentacion_health["active_presentaciones_count"] > 0 and presentacion_health["derived_presentaciones_count"] >= presentacion_health["active_presentaciones_count"] else "warning",
                "detail": (
                    "Las presentaciones activas ya quedaron sincronizadas como artículos internos."
                    if presentacion_health and presentacion_health["active_presentaciones_count"] > 0 and presentacion_health["derived_presentaciones_count"] >= presentacion_health["active_presentaciones_count"]
                    else "Todavía faltan presentaciones activas o sincronización de derivados."
                ),
                "action_label": "Administrar presentaciones",
                "action_url": reverse("recetas:presentacion_create", args=[receta.id]),
            },
            {
                "label": "Producto final",
                "status": bool(supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"]),
                "tone": "success" if supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"] else "warning",
                "detail": (
                    "Esta base ya alimenta productos finales y está integrada en la cadena comercial."
                    if supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"]
                    else "La base aún no alimenta productos finales; la cadena sigue abierta."
                ),
                "action_label": "Crear final" if not (supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"]) else "Revisar cadena",
                "action_url": f"{reverse('recetas:receta_create')}?mode=FINAL&source_base={receta.id}" if not (supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"]) else reverse("recetas:receta_detail", args=[receta.id]),
            },
            {
                "label": "MRP / Compras",
                "status": bool(supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"] and presentacion_health and presentacion_health["active_presentaciones_count"] > 0),
                "tone": "success" if supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"] and presentacion_health and presentacion_health["active_presentaciones_count"] > 0 else "warning",
                "detail": (
                    "La base ya está en condición de alimentar planeación y abastecimiento aguas abajo."
                    if supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"] and presentacion_health and presentacion_health["active_presentaciones_count"] > 0
                    else "MRP y compras solo deben usar esta base cuando derivados y consumo final queden cerrados."
                ),
                "action_label": "Abrir MRP" if supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"] else "Cerrar cadena",
                "action_url": reverse("recetas:mrp_form") if supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"] else reverse("recetas:receta_detail", args=[receta.id]),
            },
        ]
    else:
        module_enablement_cards = [
            {
                "label": "Costeo base",
                "status": bool(receta.rendimiento_cantidad and receta.rendimiento_unidad_id and cost_ready),
                "tone": "success" if receta.rendimiento_cantidad and receta.rendimiento_unidad_id and cost_ready else "warning",
                "detail": (
                    "La batida base ya tiene costo unitario estable."
                    if receta.rendimiento_cantidad and receta.rendimiento_unidad_id and cost_ready
                    else "Todavía falta rendimiento, unidad o costo para estabilizar la base."
                ),
                "action_label": "Abrir base",
                "action_url": reverse("recetas:receta_detail", args=[receta.id]),
            },
            {
                "label": "Derivados opcionales",
                "status": not receta.usa_presentaciones,
                "tone": "success" if not receta.usa_presentaciones else "warning",
                "detail": (
                    "Esta base opera sin tamaños derivados, por lo que ya puede seguir como insumo interno simple."
                    if not receta.usa_presentaciones
                    else "La base quedó marcada con presentaciones; conviene cerrarlas antes de usarla aguas abajo."
                ),
                "action_label": "Revisar base",
                "action_url": reverse("recetas:receta_detail", args=[receta.id]),
            },
            {
                "label": "Uso operativo",
                "status": master_ready and bom_ready and cost_ready,
                "tone": "success" if master_ready and bom_ready and cost_ready else "warning",
                "detail": (
                    "La base ya puede operar como insumo interno consistente."
                    if master_ready and bom_ready and cost_ready
                    else "Todavía falta cerrar maestro, estructura o costo para usar esta base con seguridad."
                ),
                "action_label": "Abrir documento",
                "action_url": reverse("recetas:receta_detail", args=[receta.id]),
            },
        ]
    module_handoff_rows = _module_enablement_handoff_rows(receta, module_enablement_cards)
    return render(
        request,
        "recetas/receta_detail.html",
        {
            "receta": receta,
            "lineas": lineas,
            "presentaciones": presentaciones,
            "total_lineas": total_lineas,
            "total_match": total_match,
            "total_revision": total_revision,
            "total_sin_match": total_sin_match,
            "total_costo_estimado": total_costo_estimado,
            "total_materia_prima": total_materia_prima,
            "total_internos": total_internos,
            "total_empaques": total_empaques,
            "total_subsecciones": total_subsecciones,
            "receta_rol_label": receta_rol_label,
            "receta_rol_description": receta_rol_description,
            "is_producto_final": is_producto_final,
            "is_base_con_presentaciones": is_base_con_presentaciones,
            "is_base_simple": is_base_simple,
            "add_line_label": add_line_label,
            "copy_label": copy_label,
            "next_steps": next_steps,
            "chain_actions": chain_actions,
            "chain_focus_summary": chain_focus_summary,
            "enterprise_stage": enterprise_stage,
            "enterprise_stage_playbook": enterprise_stage_playbook,
            "unidades": UnidadMedida.objects.order_by("codigo"),
            "tipo_choices": Receta.TIPO_CHOICES,
            "costo_por_kg_estimado": receta.costo_por_kg_estimado,
            "linea_tipo_choices": LineaReceta.TIPO_CHOICES,
            "costeo_actual": costeo_actual,
            "costeo_unavailable": costeo_unavailable,
            "versiones_recientes": versiones_recientes,
            "versiones_all": versiones_all,
            "versiones_comparativo": comparativo,
            "versiones_unavailable": versiones_unavailable,
            "selected_base": selected_base,
            "selected_target": selected_target,
            "version_compare": compare_data,
            "rend_unit_code": (receta.rendimiento_unidad.codigo if receta.rendimiento_unidad else ""),
            "familias_catalogo": familias_catalogo,
            "categorias_catalogo": categorias_catalogo,
            "familia_categoria_catalogo_json": familia_categoria_catalogo_json(categorias_catalogo),
            "recetas_copiables": recetas_copiables,
            "line_groups": line_groups,
            "recipe_master_blockers": recipe_master_blockers,
            "recipe_master_gap_totals": recipe_master_gap_totals,
            "component_breakdown": component_breakdown,
            "product_upstream_snapshot": product_upstream_snapshot,
            "derived_parent_snapshot": derived_parent_snapshot,
            "total_costo_directo": total_costo_directo,
            "bom_integrity_alerts": bom_integrity_alerts,
            "non_canonical_count": non_canonical_count,
            "direct_base_suggested_count": direct_base_suggested_count,
            "presentacion_health": presentacion_health,
            "supply_chain_snapshot": supply_chain_snapshot,
            "base_chain_actions": base_chain_actions,
            "base_chain_checkpoints": base_chain_checkpoints,
            "erp_control_chain": erp_control_chain,
            "trunk_handoff_rows": trunk_handoff_rows,
            "trunk_handoff_summary": _trunk_handoff_summary(
                trunk_handoff_rows,
                owner="Recetas / Costeo",
                fallback_url=reverse("recetas:receta_detail", args=[receta.id]),
            ),
            "module_enablement_cards": module_enablement_cards,
            "module_handoff_rows": module_handoff_rows,
            "release_gate_rows": release_gate_rows,
            "release_gate_progress": release_gate_progress,
            "erp_command_center": {
                "owner": "Recetas / Costeo",
                "status": "Crítico"
                if (total_revision + total_sin_match + recipe_master_gap_totals["total"]) > 0
                else "Seguimiento"
                if (
                    int(round((release_gate_progress["completed"] / release_gate_progress["total"]) * 100))
                    if release_gate_progress["total"]
                    else 0
                ) < 100
                else "Controlado",
                "tone": "danger"
                if (total_revision + total_sin_match + recipe_master_gap_totals["total"]) > 0
                else "warning"
                if (
                    int(round((release_gate_progress["completed"] / release_gate_progress["total"]) * 100))
                    if release_gate_progress["total"]
                    else 0
                ) < 100
                else "success",
                "blockers": total_revision + total_sin_match + recipe_master_gap_totals["total"],
                "next_step": chain_focus_summary.get("detail")
                or "Continuar el cierre del documento y su trazabilidad operativa.",
                "url": chain_focus_summary.get("action_url") or reverse("recetas:receta_detail", args=[receta.id]),
                "cta": chain_focus_summary.get("action_label") or "Abrir documento",
            },
            "erp_governance_rows": erp_governance_rows,
            "executive_radar_rows": _recipes_executive_radar_rows(
                erp_governance_rows,
                owner="Recetas / Costeo",
                fallback_url=reverse("recetas:receta_detail", args=[receta.id]),
            ),
            "critical_path_rows": _recipes_critical_path_rows(
                erp_governance_rows,
                owner="Recetas / Costeo",
                fallback_url=reverse("recetas:receta_detail", args=[receta.id]),
            ),
        },
    )


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def receta_update(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    nombre = (request.POST.get("nombre") or "").strip()
    codigo_point = (request.POST.get("codigo_point") or "").strip()
    familia = (request.POST.get("familia") or "").strip()
    categoria = (request.POST.get("categoria") or "").strip()
    sheet_name = (request.POST.get("sheet_name") or "").strip()
    tipo = (request.POST.get("tipo") or Receta.TIPO_PREPARACION).strip()
    usa_presentaciones = request.POST.get("usa_presentaciones") == "on"
    rendimiento_cantidad = _to_decimal_or_none(request.POST.get("rendimiento_cantidad"))
    rendimiento_unidad_id = request.POST.get("rendimiento_unidad_id")

    if not nombre:
        messages.error(request, "El nombre de receta es obligatorio.")
        return redirect("recetas:receta_detail", pk=pk)

    if tipo not in {Receta.TIPO_PREPARACION, Receta.TIPO_PRODUCTO_FINAL}:
        tipo = Receta.TIPO_PREPARACION

    validation_errors = _validate_receta_enterprise_rules(
        tipo=tipo,
        usa_presentaciones=usa_presentaciones,
        rendimiento_cantidad=rendimiento_cantidad,
        rendimiento_unidad_id=rendimiento_unidad_id,
        familia=familia,
        presentaciones_existentes=receta.presentaciones.count(),
    )
    if validation_errors:
        for error in validation_errors:
            messages.error(request, error)
        return redirect("recetas:receta_detail", pk=pk)

    if tipo == Receta.TIPO_PRODUCTO_FINAL:
        usa_presentaciones = False
        rendimiento_cantidad = None
        rendimiento_unidad_id = ""

    receta.nombre = nombre[:250]
    receta.codigo_point = codigo_point[:80]
    receta.familia = familia[:120]
    receta.categoria = categoria[:120]
    receta.sheet_name = sheet_name[:120]
    receta.tipo = tipo
    receta.usa_presentaciones = usa_presentaciones
    receta.rendimiento_cantidad = rendimiento_cantidad
    receta.rendimiento_unidad = UnidadMedida.objects.filter(pk=rendimiento_unidad_id).first() if rendimiento_unidad_id else None
    receta.save()
    _sync_derived_insumos_safe(request, receta)
    _sync_cost_version_safe(request, receta, "RECETA_UPDATE")
    messages.success(request, "Receta actualizada.")
    return redirect("recetas:receta_detail", pk=pk)


def _receta_delete_blockers(receta: Receta) -> list[tuple[str, int]]:
    blockers = [
        ("ventas históricas", receta.ventas_historicas.count()),
        ("ventas POS", receta.ventas_pos.count()),
        ("mermas POS", receta.mermas_pos.count()),
        ("solicitudes de venta", receta.solicitudes_venta.count()),
        ("pronósticos", receta.pronosticos.count()),
        ("renglones de plan", receta.plan_items.count()),
        ("políticas de stock por sucursal", receta.politicas_stock_sucursal.count()),
        ("movimientos CEDIS", receta.movimientos_cedis.count()),
        ("líneas de reabasto CEDIS", receta.solicitudes_reabasto_lineas.count()),
        ("reservas pickup", receta.pickup_reservations.count()),
        ("ventas diarias Point", receta.point_daily_sales.count()),
    ]
    if getattr(receta, "inventario_cedis", None):
        blockers.append(("inventario CEDIS", 1))
    return [(label, count) for label, count in blockers if count]


@login_required
@permission_required("recetas.delete_receta", raise_exception=True)
@require_POST
def receta_delete(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    blockers = _receta_delete_blockers(receta)
    if blockers:
        summary = ", ".join(f"{count} {label}" for label, count in blockers[:4])
        extra = f" y {len(blockers) - 4} más" if len(blockers) > 4 else ""
        messages.error(
            request,
            f"No se puede eliminar la receta porque ya tiene huella operativa: {summary}{extra}.",
        )
        return redirect("recetas:receta_detail", pk=pk)

    nombre = receta.nombre
    receta_id = receta.id
    try:
        receta.delete()
    except ProtectedError:
        messages.error(
            request,
            "No se puede eliminar la receta porque todavía está ligada a documentos operativos protegidos.",
        )
        return redirect("recetas:receta_detail", pk=pk)

    log_event(
        request.user,
        "DELETE",
        "recetas.Receta",
        str(receta_id),
        {"receta_id": receta_id, "nombre": nombre},
    )
    messages.success(request, f"Receta eliminada: {nombre}.")
    return redirect("recetas:recetas_list")


def _to_decimal_or_none(value: str | None) -> Decimal | None:
    if value is None:
        return None
    raw = str(value).strip().replace(",", ".")
    if raw == "":
        return None
    try:
        return Decimal(raw)
    except Exception:
        return None


def _to_non_negative_decimal_or_none(value: str | None) -> Decimal | None:
    if value is None:
        return None
    raw = str(value).strip().replace(",", ".")
    if raw == "":
        return None
    try:
        num = Decimal(raw)
        return num if num >= 0 else None
    except Exception:
        return None


def _validate_receta_enterprise_rules(
    *,
    mode_code: str | None = None,
    tipo: str | None = None,
    usa_presentaciones: bool = False,
    rendimiento_cantidad: Decimal | None = None,
    rendimiento_unidad_id: str | int | None = None,
    familia: str = "",
    presentaciones_existentes: int = 0,
) -> list[str]:
    errors: list[str] = []
    normalized_mode = (mode_code or "").strip().upper()
    receta_tipo = (tipo or Receta.TIPO_PREPARACION).strip()
    is_producto_final = normalized_mode == "FINAL" or receta_tipo == Receta.TIPO_PRODUCTO_FINAL
    is_base_con_presentaciones = normalized_mode == "BASE_DERIVADOS" or (
        receta_tipo == Receta.TIPO_PREPARACION and usa_presentaciones
    )
    is_base_simple = normalized_mode == "BASE" or (
        receta_tipo == Receta.TIPO_PREPARACION and not usa_presentaciones
    )

    has_rendimiento = bool(rendimiento_cantidad and rendimiento_cantidad > 0)
    has_rendimiento_unidad = bool(str(rendimiento_unidad_id or "").strip())

    if is_producto_final:
        if not familia:
            errors.append("En producto final debes seleccionar una familia.")
        if usa_presentaciones:
            errors.append("Un producto final no puede usar presentaciones derivadas.")
        if presentaciones_existentes:
            errors.append(
                "No puedes convertir a producto final una receta que aún tiene presentaciones activas. Elimínalas o consolídalas primero."
            )
        return errors

    if is_base_simple and presentaciones_existentes:
        errors.append(
            "No puedes convertir a base simple una receta que ya tiene presentaciones activas. Elimínalas primero o conserva el modo con derivados."
        )

    if is_base_simple or is_base_con_presentaciones:
        if not has_rendimiento:
            errors.append("Debes capturar el rendimiento total de la batida para costeo enterprise.")
        if not has_rendimiento_unidad:
            errors.append("Debes seleccionar la unidad del rendimiento para costeo enterprise.")

    return errors


def _validate_presentacion_workflow(receta: Receta) -> list[str]:
    errors: list[str] = []
    if receta.tipo != Receta.TIPO_PREPARACION:
        errors.append("Solo un insumo base permite presentaciones derivadas.")
        return errors
    if not receta.rendimiento_cantidad or receta.rendimiento_cantidad <= 0:
        errors.append("Antes de administrar presentaciones debes capturar el rendimiento total de la base.")
    if not receta.rendimiento_unidad_id:
        errors.append("Antes de administrar presentaciones debes capturar la unidad del rendimiento.")
    return errors


def _linea_form_context(
    receta: Receta,
    linea: LineaReceta | None = None,
    *,
    advanced_mode: bool = False,
    component_filter: str | None = None,
    component_context: str | None = None,
) -> Dict[str, Any]:
    raw_insumos = [
        insumo
        for insumo in canonicalized_insumo_selector(limit=1600)
        if insumo.activo and insumo.unidad_base_id
    ]
    selector_source_recipe_ids = {
        recipe_id
        for insumo in raw_insumos
        for recipe_id in [_recipe_id_from_derived_code(insumo.codigo or "")]
        if recipe_id
    }
    selector_source_recipe_map = {
        item.id: item
        for item in Receta.objects.filter(id__in=selector_source_recipe_ids).only("id", "usa_presentaciones")
    }
    selector_source_presentaciones = {
        row["receta_id"]: row["total"]
        for row in (
            RecetaPresentacion.objects.filter(receta_id__in=selector_source_recipe_ids, activo=True)
            .values("receta_id")
            .annotate(total=Count("id"))
        )
    }

    generic_recipe_tokens = {
        "pastel",
        "pasteles",
        "producto",
        "final",
        "base",
        "batida",
        "mezcla",
        "chico",
        "mediano",
        "grande",
        "mini",
        "bollos",
        "bollo",
        "individual",
        "rosca",
        "media",
        "plancha",
        "de",
        "del",
        "con",
        "para",
        "y",
    }
    recipe_context_tokens = [
        token
        for token in normalizar_nombre(f"{receta.nombre} {receta.familia or ''} {receta.categoria or ''}").split()
        if len(token) >= 3 and token not in generic_recipe_tokens
    ]

    def _recipe_context_score(i: Insumo) -> int:
        if not recipe_context_tokens:
            return 0
        haystack = normalizar_nombre(" ".join([i.nombre or "", i.categoria or "", i.codigo or "", i.codigo_point or ""]))
        score = 0
        for token in recipe_context_tokens:
            if token and token in haystack:
                score += 1
        return score

    def _option_score(i: Insumo) -> int:
        score = 0
        code = (i.codigo or "")
        derived_kind = _derived_code_kind(code)
        source_recipe_id = _recipe_id_from_derived_code(code)
        source_recipe = selector_source_recipe_map.get(source_recipe_id) if source_recipe_id else None
        source_active_presentaciones = int(selector_source_presentaciones.get(source_recipe_id, 0)) if source_recipe_id else 0
        is_internal = i.tipo_item == Insumo.TIPO_INTERNO or code.startswith("DERIVADO:RECETA:")
        is_empaque = i.tipo_item == Insumo.TIPO_EMPAQUE or "empaque" in normalizar_nombre(i.categoria or "")
        if is_internal:
            score += 300
            if derived_kind == "PREPARACION":
                score += 50
            elif derived_kind == "PRESENTACION":
                score += 120
        elif is_empaque:
            score += 120
        if (
            receta.tipo == Receta.TIPO_PRODUCTO_FINAL
            and derived_kind == "PREPARACION"
            and source_recipe
            and source_recipe.usa_presentaciones
            and source_active_presentaciones > 0
        ):
            score -= 240
        if (
            receta.tipo == Receta.TIPO_PRODUCTO_FINAL
            and derived_kind == "PRESENTACION"
            and source_recipe
            and source_recipe.usa_presentaciones
            and source_active_presentaciones > 0
        ):
            score += 180
        if (i.codigo_point or "").strip():
            score += 80
        if i.latest_costo_unitario is not None and Decimal(str(i.latest_costo_unitario or 0)) > 0:
            score += 60
        erp_profile = getattr(i, "erp_profile", None) or _insumo_erp_readiness(i)
        if erp_profile["ready"]:
            score += 140
        else:
            score -= 160
        score += _recipe_context_score(i) * 45
        if i.proveedor_principal_id:
            score += 20
        if i.unidad_base_id:
            score += 10
        # Evita que "ruido" de pruebas prevalezca sobre catálogo real.
        if normalizar_nombre(i.nombre).startswith("qa flow"):
            score -= 120
        return score

    for insumo in raw_insumos:
        if not hasattr(insumo, "origen_orden"):
            insumo.origen_orden = (
                0
                if insumo.tipo_item == Insumo.TIPO_INTERNO or (insumo.codigo or "").startswith("DERIVADO:RECETA:")
                else 1
                if insumo.tipo_item == Insumo.TIPO_EMPAQUE
                else 2
            )
        if not hasattr(insumo, "canonical_variant_count"):
            insumo.canonical_variant_count = len(getattr(insumo, "member_ids", []) or [insumo.id])
        insumo.recipe_context_score = _recipe_context_score(insumo)
        insumo.erp_profile = _insumo_erp_readiness(insumo)
        insumo.source_recipe_id = _recipe_id_from_derived_code(insumo.codigo or "")
        insumo.source_code_kind = _derived_code_kind(insumo.codigo or "")
        insumo.source_active_presentaciones = int(
            selector_source_presentaciones.get(insumo.source_recipe_id, 0)
        ) if insumo.source_recipe_id else 0
        insumo.option_score = _option_score(insumo)

    insumos = sorted(
        raw_insumos,
        key=lambda x: ((x.origen_orden or 9), -(x.option_score or 0), x.nombre.lower()),
    )
    if linea and linea.insumo_id:
        selected_present = any(i.id == linea.insumo_id for i in insumos)
        if not selected_present and linea.insumo and linea.insumo.activo and linea.insumo.unidad_base_id:
            linea.insumo.erp_profile = _insumo_erp_readiness(linea.insumo)
            linea.insumo.option_score = _option_score(linea.insumo)
            insumos.append(linea.insumo)
            insumos = sorted(
                insumos,
                key=lambda x: ((x.origen_orden or 9), -(getattr(x, "option_score", 0) or 0), x.nombre.lower()),
            )

    def _is_internal(i: Insumo) -> bool:
        return i.tipo_item == Insumo.TIPO_INTERNO or (i.codigo or "").startswith("DERIVADO:RECETA:")

    def _is_empaque(i: Insumo) -> bool:
        return i.tipo_item == Insumo.TIPO_EMPAQUE or "empaque" in normalizar_nombre(i.categoria or "")

    insumos_internos = [i for i in insumos if _is_internal(i)]
    insumos_empaque = [i for i in insumos if not _is_internal(i) and _is_empaque(i)]
    insumos_materia_prima = [i for i in insumos if not _is_internal(i) and not _is_empaque(i)]
    insumos_internos_ready_count = sum(1 for i in insumos_internos if i.erp_profile["ready"])
    insumos_empaque_ready_count = sum(1 for i in insumos_empaque if i.erp_profile["ready"])
    insumos_materia_prima_ready_count = sum(1 for i in insumos_materia_prima if i.erp_profile["ready"])
    quick_mode = receta.tipo == Receta.TIPO_PRODUCTO_FINAL and not advanced_mode and (
        linea is None or linea.tipo_linea == LineaReceta.TIPO_NORMAL
    )
    if component_filter not in {Insumo.TIPO_INTERNO, Insumo.TIPO_MATERIA_PRIMA, Insumo.TIPO_EMPAQUE, "ALL"}:
        component_filter = Insumo.TIPO_INTERNO if receta.tipo == Receta.TIPO_PRODUCTO_FINAL else Insumo.TIPO_MATERIA_PRIMA
    component_filter_meta = {
        "ALL": {
            "title": "Todo el catálogo",
            "description": "Muestra insumos internos, materia prima y empaques. Úsalo solo cuando no estés seguro de la clase.",
            "count": len(insumos),
        },
        Insumo.TIPO_INTERNO: {
            "title": "Insumos internos",
            "description": "Panes, rellenos, coberturas, batidas y subinsumos producidos dentro de la empresa.",
            "count": len(insumos_internos),
        },
        Insumo.TIPO_MATERIA_PRIMA: {
            "title": "Materia prima",
            "description": "Artículos comprados directo a proveedor y usados tal cual en la receta o batida.",
            "count": len(insumos_materia_prima),
        },
        Insumo.TIPO_EMPAQUE: {
            "title": "Empaques",
            "description": "Domo, caja, etiqueta, vaso y cualquier material de presentación final.",
            "count": len(insumos_empaque),
        },
    }
    component_context_map = {
        "internos": {
            "title": "Bloque: insumos internos",
            "description": "Estás agregando un componente producido dentro de la empresa: pan, relleno, cobertura, batida o subinsumo.",
        },
        "materia_prima": {
            "title": "Bloque: materia prima puntual",
            "description": "Estás agregando un artículo que entra directo al armado final sin pasar por una batida interna.",
        },
        "empaques": {
            "title": "Bloque: empaque final",
            "description": "Estás agregando caja, domo, etiqueta, vaso u otro material de presentación final.",
        },
        "subsecciones": {
            "title": "Bloque: subsección operativa",
            "description": "Estás capturando un bloque lógico como relleno, decorado o cobertura para ordenar el armado.",
        },
        "general": {
            "title": "Bloque general",
            "description": "Captura el componente dentro del bloque correcto para mantener el BOM ordenado.",
        },
    }
    if component_context not in component_context_map:
        component_context = "general"
    show_subsection_controls = component_context == "subsecciones" or (
        linea is not None and linea.tipo_linea == LineaReceta.TIPO_SUBSECCION
    )
    component_filter_choices = ["ALL", Insumo.TIPO_INTERNO, Insumo.TIPO_MATERIA_PRIMA, Insumo.TIPO_EMPAQUE]
    if quick_mode:
        component_filter_choices = [Insumo.TIPO_INTERNO, Insumo.TIPO_MATERIA_PRIMA, Insumo.TIPO_EMPAQUE]
    return {
        "receta": receta,
        "linea": linea,
        "advanced_mode": advanced_mode,
        "quick_mode": quick_mode,
        "default_component_filter": component_filter,
        "insumos": insumos,
        "insumos_internos": insumos_internos,
        "insumos_empaque": insumos_empaque,
        "insumos_materia_prima": insumos_materia_prima,
        "insumos_internos_count": len(insumos_internos),
        "insumos_empaque_count": len(insumos_empaque),
        "insumos_materia_prima_count": len(insumos_materia_prima),
        "insumos_internos_ready_count": insumos_internos_ready_count,
        "insumos_empaque_ready_count": insumos_empaque_ready_count,
        "insumos_materia_prima_ready_count": insumos_materia_prima_ready_count,
        "component_filter_meta": component_filter_meta,
        "component_context": component_context,
        "component_context_meta": component_context_map[component_context],
        "show_subsection_controls": show_subsection_controls,
        "component_filter_choices": component_filter_choices,
        "product_final_component_guidance": [
            {
                "filter": "INSUMO_INTERNO",
                "title": "Insumo interno",
                "description": "Usa panes, rellenos, coberturas y subinsumos ya producidos dentro de la empresa.",
            },
            {
                "filter": "MATERIA_PRIMA",
                "title": "Materia prima puntual",
                "description": "Úsala solo cuando el artículo entra directo al armado final y no existe como insumo interno.",
            },
            {
                "filter": "EMPAQUE",
                "title": "Empaque",
                "description": "Úsalo para domos, cajas, etiquetas, vasos y demás material de presentación final.",
            },
        ],
        "unidades": UnidadMedida.objects.order_by("codigo"),
        "linea_tipo_choices": LineaReceta.TIPO_CHOICES,
    }


def _latest_cost_for_insumo(insumo: Insumo | None) -> Decimal | None:
    if not insumo:
        return None
    cost, _unit, _source = resolve_insumo_unit_cost(insumo)
    return cost


def _switch_line_to_internal_cost(linea: LineaReceta) -> None:
    # Si ya hay cantidad + insumo, dejamos de usar costo fijo de Excel y
    # pasamos a costo dinámico interno (cantidad * costo_unitario_snapshot).
    if not linea.insumo:
        return
    if linea.cantidad is None or linea.cantidad <= 0:
        return

    if linea.costo_unitario_snapshot is None or linea.costo_unitario_snapshot <= 0:
        latest, _source = resolve_line_snapshot_cost(linea)
        if latest is not None and latest > 0:
            linea.costo_unitario_snapshot = latest
        elif linea.costo_linea_excel is not None and linea.costo_linea_excel > 0:
            try:
                linea.costo_unitario_snapshot = linea.costo_linea_excel / linea.cantidad
            except Exception:
                pass

    if linea.costo_unitario_snapshot is not None and linea.costo_unitario_snapshot > 0:
        linea.costo_linea_excel = None


def _autofill_unidad_from_insumo(linea: LineaReceta) -> None:
    if not linea.insumo:
        return
    if linea.unidad is None and linea.insumo.unidad_base is not None:
        linea.unidad = linea.insumo.unidad_base
    if linea.insumo.unidad_base is not None:
        # La unidad del componente ligado es fija para evitar inconsistencias de costeo.
        linea.unidad = linea.insumo.unidad_base
    if linea.unidad and not (linea.unidad_texto or "").strip():
        linea.unidad_texto = linea.unidad.codigo
    if linea.unidad:
        linea.unidad_texto = linea.unidad.codigo


def _validate_linea_operativa(receta: Receta, linea: LineaReceta) -> str | None:
    if linea.tipo_linea != LineaReceta.TIPO_NORMAL:
        return None
    if receta.tipo == Receta.TIPO_PRODUCTO_FINAL and not linea.insumo_id:
        return "En producto final, cada renglón principal debe estar ligado a un insumo/subinsumo."
    if linea.insumo_id and linea.insumo and linea.insumo.unidad_base_id is None:
        return (
            f"El insumo ligado ({linea.insumo.nombre}) no tiene unidad base. "
            "Asigna unidad en Catálogo de insumos para poder costear."
        )
    if linea.insumo_id and (linea.cantidad is None or linea.cantidad <= 0):
        unidad = linea.unidad.codigo if linea.unidad else (linea.unidad_texto or "").strip()
        if unidad:
            return f"Captura cantidad mayor a cero para el insumo ligado ({unidad})."
        return "Captura cantidad mayor a cero para el insumo ligado."

    # Regla operativa: en recetas tipo bollo, el pan se consume desde batida base (kg),
    # no desde presentaciones en pieza.
    presentacion = _extract_presentacion_from_recipe_name(receta.nombre)
    ingredient_text = ((linea.insumo_texto or "").strip() or (linea.insumo.nombre if linea.insumo else "")).lower()
    if presentacion in {"Bollo", "Bollos", "Bollito"} and ingredient_text.startswith("pan"):
        unit = linea.insumo.unidad_base if (linea.insumo and linea.insumo.unidad_base_id) else None
        if not unit or unit.tipo != UnidadMedida.TIPO_MASA:
            return (
                "En recetas de bollo, el pan debe ligarse a la batida base en kg. "
                "Selecciona el insumo derivado de preparación (no presentación por pieza)."
            )
    return None


def _line_component_kind(insumo: Insumo | None) -> str | None:
    if not insumo:
        return None
    if insumo.tipo_item == Insumo.TIPO_EMPAQUE:
        return Insumo.TIPO_EMPAQUE
    if insumo.tipo_item == Insumo.TIPO_INTERNO or (insumo.codigo or "").startswith("DERIVADO:RECETA:"):
        return Insumo.TIPO_INTERNO
    return Insumo.TIPO_MATERIA_PRIMA


def _validate_component_filter_selection(
    receta: Receta,
    linea: LineaReceta,
    component_filter: str | None,
) -> str | None:
    if receta.tipo != Receta.TIPO_PRODUCTO_FINAL:
        return None
    if not linea.insumo_id or not component_filter or component_filter == "ALL":
        return None
    expected = component_filter.strip().upper()
    if expected not in {Insumo.TIPO_INTERNO, Insumo.TIPO_MATERIA_PRIMA, Insumo.TIPO_EMPAQUE}:
        return None

    actual = _line_component_kind(linea.insumo)
    if actual == expected:
        return None

    expected_labels = {
        Insumo.TIPO_INTERNO: "insumo interno",
        Insumo.TIPO_MATERIA_PRIMA: "materia prima puntual",
        Insumo.TIPO_EMPAQUE: "empaque",
    }
    actual_labels = {
        Insumo.TIPO_INTERNO: "insumo interno",
        Insumo.TIPO_MATERIA_PRIMA: "materia prima",
        Insumo.TIPO_EMPAQUE: "empaque",
    }
    return (
        f"Seleccionaste el flujo de {expected_labels[expected]}, pero el artículo ligado "
        f"pertenece a {actual_labels.get(actual, 'otra clase')}. "
        "Cambia la clase de componente o selecciona un artículo del grupo correcto."
    )


def _validate_quick_mode_selection(
    receta: Receta,
    linea: LineaReceta,
    quick_mode: bool,
) -> str | None:
    if not quick_mode:
        return None
    if receta.tipo != Receta.TIPO_PRODUCTO_FINAL:
        return None
    if linea.tipo_linea != LineaReceta.TIPO_NORMAL:
        return None
    if not linea.insumo_id:
        return "En producto final, el modo rápido requiere seleccionar un artículo estándar del catálogo antes de guardar."
    latest_cost = _latest_cost_for_insumo(linea.insumo)
    if latest_cost is None or latest_cost <= 0:
        return (
            f"El artículo seleccionado ({linea.insumo.nombre}) no tiene costo vigente. "
            "Primero registra costo en el maestro para poder usarlo en producto final."
        )
    return None


def _repoint_linea_to_canonical_if_needed(linea: LineaReceta) -> bool:
    if not linea.insumo_id:
        return False
    canonical = _canonicalize_insumo_match(linea.insumo)
    if not canonical or canonical.id == linea.insumo_id:
        return False
    linea.insumo = canonical
    linea.insumo_texto = _insumo_display_name(canonical)[:250]
    return True


def _apply_direct_base_replacement_to_line(
    linea: LineaReceta,
    replacement: dict[str, object],
) -> None:
    linea.insumo = replacement["insumo"]
    linea.insumo_texto = _insumo_display_name(replacement["insumo"])
    linea.unidad = replacement["insumo"].unidad_base
    linea.unidad_texto = (
        replacement["insumo"].unidad_base.codigo
        if replacement["insumo"].unidad_base_id
        else linea.unidad_texto
    )
    linea.cantidad = replacement["replacement_quantity"]
    linea.costo_unitario_snapshot = None
    linea.costo_linea_excel = None
    linea.match_status = LineaReceta.STATUS_AUTO
    linea.match_method = "DIRECT_BASE_PRESENTACION"
    linea.match_score = 100


def _extract_presentacion_from_recipe_name(recipe_name: str) -> str | None:
    normalized = normalizar_nombre(recipe_name or "")
    if not normalized:
        return None
    token_map = [
        ("media plancha", "Media Plancha"),
        ("1 2 plancha", "Media Plancha"),
        ("rosca", "Rosca"),
        ("individual", "Individual"),
        ("bollito", "Bollito"),
        ("bollos", "Bollos"),
        ("bollo", "Bollos"),
        ("grande", "Grande"),
        ("mediano", "Mediano"),
        ("chico", "Chico"),
        ("mini", "Mini"),
    ]
    for token, label in token_map:
        if f" {token}" in f" {normalized}" or normalized.endswith(token):
            return label
    return None


def _meaningful_pan_tokens(*parts: str) -> list[str]:
    generic_tokens = {
        "pan",
        "de",
        "del",
        "la",
        "el",
        "los",
        "las",
        "con",
        "para",
        "y",
        "pastel",
        "pasteles",
        "producto",
        "final",
        "base",
        "batida",
        "mezcla",
        "relleno",
        "cobertura",
        "decorado",
        "decoracion",
        "chico",
        "mediano",
        "grande",
        "mini",
        "individual",
        "bollo",
        "bollos",
        "bollito",
        "rosca",
        "media",
        "plancha",
    }
    tokens: list[str] = []
    for part in parts:
        for token in normalizar_nombre(part or "").split():
            if len(token) < 3 or token in generic_tokens:
                continue
            if token not in tokens:
                tokens.append(token)
    return tokens


def _select_best_pan_derived_candidate(
    recipe_name: str,
    ingredient_text: str,
) -> Insumo | None:
    presentacion = _extract_presentacion_from_recipe_name(recipe_name)
    if not presentacion:
        return None

    ingredient_norm = normalizar_nombre(ingredient_text or "")
    recipe_norm = normalizar_nombre(recipe_name or "")
    full_context = f"{ingredient_norm} {recipe_norm}".strip()
    if not full_context.startswith("pan "):
        return None
    if full_context.startswith("decorado pan"):
        return None
    flavor_tokens = _meaningful_pan_tokens(ingredient_text, recipe_name)
    if not flavor_tokens:
        return None

    is_bollo_recipe = presentacion in {"Bollo", "Bollos", "Bollito"}
    presentacion_norm = normalizar_nombre(presentacion or "")
    candidates = list(
        Insumo.objects.filter(codigo__startswith="DERIVADO:RECETA:", activo=True, nombre__icontains="Pan")
        .only("id", "nombre", "codigo", "unidad_base_id")
        .order_by("nombre", "id")
    )
    best_candidate: Insumo | None = None
    best_score = -1
    for candidate in candidates:
        candidate_norm = normalizar_nombre(candidate.nombre or "")
        token_hits = [token for token in flavor_tokens if token in candidate_norm]
        if not token_hits:
            continue
        derived_kind = _derived_code_kind(candidate.codigo or "")
        score = len(token_hits) * 100
        if presentacion_norm and derived_kind == "PRESENTACION" and presentacion_norm in candidate_norm:
            score += 90
        if is_bollo_recipe and derived_kind == "PREPARACION":
            score += 140
        elif is_bollo_recipe and derived_kind == "PRESENTACION":
            score -= 120
        elif not is_bollo_recipe and derived_kind == "PRESENTACION":
            score += 50
        elif not is_bollo_recipe and derived_kind == "PREPARACION":
            score -= 40
        if candidate_norm.startswith("pan "):
            score += 10
        if score > best_score:
            best_score = score
            best_candidate = candidate
    return best_candidate if best_score >= 100 else None


def _autolink_pan_derived_from_recipe(receta: Receta, linea: LineaReceta) -> None:
    ingredient_text = (linea.insumo_texto or "").strip()
    if not ingredient_text and linea.insumo:
        ingredient_text = _insumo_display_name(linea.insumo)
    candidate = _select_best_pan_derived_candidate(receta.nombre, ingredient_text)
    if not candidate:
        return
    if linea.insumo_id == candidate.id:
        return
    linea.insumo = candidate


def _sync_derived_insumos_safe(request: HttpRequest, receta: Receta) -> None:
    try:
        sync_receta_derivados(receta)
    except Exception:
        messages.warning(
            request,
            "La receta se guardó, pero falló la sincronización automática de costos/insumos derivados.",
        )


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
def receta_sync_derivados(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    _sync_derived_insumos_safe(request, receta)
    _sync_cost_version_safe(request, receta, "SYNC_DERIVADOS_MANUAL")
    messages.success(request, "Derivados actualizados.")
    next_url = (request.GET.get("next") or "").strip()
    if next_url:
        return redirect(next_url)
    return redirect("recetas:receta_detail", pk=pk)


def _sync_cost_version_safe(request: HttpRequest, receta: Receta, fuente: str) -> None:
    try:
        asegurar_version_costeo(receta, fuente=fuente)
    except Exception:
        messages.warning(
            request,
            "Se guardaron cambios, pero falló el versionado automático de costos.",
        )


def _next_line_position(receta: Receta) -> int:
    last = receta.lineas.order_by("-posicion").values_list("posicion", flat=True).first()
    return int(last or 0) + 1


def _clone_linea_to_receta(
    source_linea: LineaReceta,
    target_receta: Receta,
    position: int,
) -> LineaReceta:
    return LineaReceta.objects.create(
        receta=target_receta,
        posicion=position,
        tipo_linea=source_linea.tipo_linea,
        etapa=source_linea.etapa,
        insumo=source_linea.insumo,
        insumo_texto=source_linea.insumo_texto,
        cantidad=source_linea.cantidad,
        unidad_texto=source_linea.unidad_texto,
        unidad=source_linea.unidad,
        costo_linea_excel=source_linea.costo_linea_excel,
        costo_unitario_snapshot=source_linea.costo_unitario_snapshot,
        match_score=source_linea.match_score,
        match_method=source_linea.match_method,
        match_status=source_linea.match_status,
        aprobado_por=source_linea.aprobado_por,
        aprobado_en=source_linea.aprobado_en,
    )


def _resequence_lineas(receta: Receta) -> None:
    changed: list[LineaReceta] = []
    for idx, linea in enumerate(receta.lineas.order_by("posicion", "id"), start=1):
        if linea.posicion != idx:
            linea.posicion = idx
            changed.append(linea)
    if changed:
        LineaReceta.objects.bulk_update(changed, ["posicion"])


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def receta_copy_lineas(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    source_id = (request.POST.get("source_receta_id") or "").strip()
    copy_mode = (request.POST.get("copy_mode") or "append").strip().lower()

    if not source_id.isdigit():
        messages.error(request, "Selecciona la receta origen que quieres copiar.")
        return redirect("recetas:receta_detail", pk=pk)

    receta_origen = get_object_or_404(Receta, pk=int(source_id))
    if receta_origen.pk == receta.pk:
        messages.error(request, "No puedes copiar ingredientes desde la misma receta.")
        return redirect("recetas:receta_detail", pk=pk)

    if copy_mode not in {"append", "replace"}:
        copy_mode = "append"

    lineas_origen = list(receta_origen.lineas.order_by("posicion", "id"))
    if not lineas_origen:
        messages.error(request, "La receta origen no tiene ingredientes para copiar.")
        return redirect("recetas:receta_detail", pk=pk)

    current_count = receta.lineas.count()
    deleted_count = 0
    copied_count = 0

    with transaction.atomic():
        if copy_mode == "replace":
            deleted_count, _ = receta.lineas.all().delete()
            start_position = 1
        else:
            start_position = _next_line_position(receta)

        position = start_position
        for source_linea in lineas_origen:
            _clone_linea_to_receta(source_linea, receta, position)
            position += 1
            copied_count += 1

        _resequence_lineas(receta)

    _sync_derived_insumos_safe(request, receta)
    _sync_cost_version_safe(request, receta, "LINEAS_COPY")
    log_event(
        request.user,
        "COPY_LINEAS",
        "recetas.Receta",
        str(receta.pk),
        {
            "receta_destino_id": receta.pk,
            "receta_destino_nombre": receta.nombre,
            "receta_origen_id": receta_origen.pk,
            "receta_origen_nombre": receta_origen.nombre,
            "copy_mode": copy_mode,
            "copied_count": copied_count,
            "deleted_count": deleted_count,
            "lineas_previas_destino": current_count,
        },
    )

    if copy_mode == "replace":
        messages.success(
            request,
            f"Se reemplazaron {deleted_count} líneas y se copiaron {copied_count} ingredientes desde {receta_origen.nombre}.",
        )
    else:
        messages.success(
            request,
            f"Se copiaron {copied_count} ingredientes desde {receta_origen.nombre} al final de la receta.",
        )
    return redirect("recetas:receta_detail", pk=pk)


def _load_versiones_costeo(receta: Receta, limit: int) -> list[RecetaCostoVersion]:
    return list(receta.versiones_costo.order_by("-version_num")[:limit])


def _empty_costeo_actual() -> dict[str, Any]:
    return {
        "driver": None,
        "costo_mp": Decimal("0"),
        "costo_mo": Decimal("0"),
        "costo_indirecto": Decimal("0"),
        "costo_total": Decimal("0"),
        "costo_por_unidad_rendimiento": Decimal("0"),
    }


def _compare_versions(base: RecetaCostoVersion, target: RecetaCostoVersion) -> dict[str, Decimal]:
    delta_mp = Decimal(str(target.costo_mp or 0)) - Decimal(str(base.costo_mp or 0))
    delta_mo = Decimal(str(target.costo_mo or 0)) - Decimal(str(base.costo_mo or 0))
    delta_ind = Decimal(str(target.costo_indirecto or 0)) - Decimal(str(base.costo_indirecto or 0))
    delta_total = Decimal(str(target.costo_total or 0)) - Decimal(str(base.costo_total or 0))

    base_total = Decimal(str(base.costo_total or 0))
    delta_pct_total = None
    if base_total > 0:
        delta_pct_total = (delta_total * Decimal("100")) / base_total

    base_unidad = Decimal(str(base.costo_por_unidad_rendimiento or 0))
    target_unidad = Decimal(str(target.costo_por_unidad_rendimiento or 0))
    delta_unidad = target_unidad - base_unidad
    delta_pct_unidad = None
    if base_unidad > 0:
        delta_pct_unidad = (delta_unidad * Decimal("100")) / base_unidad

    return {
        "delta_mp": delta_mp,
        "delta_mo": delta_mo,
        "delta_indirecto": delta_ind,
        "delta_total": delta_total,
        "delta_pct_total": delta_pct_total,
        "delta_unidad": delta_unidad,
        "delta_pct_unidad": delta_pct_unidad,
    }


def _map_driver_header(header: str) -> str:
    key = normalizar_nombre(header).replace("_", " ")
    if key in {"scope", "alcance", "tipo"}:
        return "scope"
    if key in {"nombre", "driver", "driver nombre"}:
        return "nombre"
    if key in {"receta", "producto", "nombre receta"}:
        return "receta"
    if key in {"codigo point", "codigo", "sku"}:
        return "codigo_point"
    if key in {"familia", "sheet", "categoria"}:
        return "familia"
    if key in {"lote desde", "lote min", "desde"}:
        return "lote_desde"
    if key in {"lote hasta", "lote max", "hasta"}:
        return "lote_hasta"
    if key in {"mo pct", "mo%", "mano obra pct", "mano de obra pct"}:
        return "mo_pct"
    if key in {"ind pct", "indirecto pct", "indirectos pct", "indirecto%"}:
        return "indirecto_pct"
    if key in {"mo fijo", "mano obra fijo", "mano de obra fijo"}:
        return "mo_fijo"
    if key in {"ind fijo", "indirecto fijo", "indirectos fijo"}:
        return "indirecto_fijo"
    if key in {"prioridad", "priority"}:
        return "prioridad"
    if key in {"activo", "enabled"}:
        return "activo"
    return key


def _normalize_driver_scope(raw: str | None) -> str:
    key = normalizar_nombre(raw or "")
    if key in {"producto", "product"}:
        return CostoDriver.SCOPE_PRODUCTO
    if key in {"familia", "family"}:
        return CostoDriver.SCOPE_FAMILIA
    if key in {"lote", "batch"}:
        return CostoDriver.SCOPE_LOTE
    return CostoDriver.SCOPE_GLOBAL


def _to_int_safe(value: object, default: int = 0) -> int:
    try:
        return int(str(value).strip())
    except Exception:
        return default


def _to_bool_safe(value: object, default: bool = True) -> bool:
    if value is None:
        return default
    key = normalizar_nombre(str(value))
    if key in {"1", "true", "si", "yes", "on", "activo"}:
        return True
    if key in {"0", "false", "no", "off", "inactivo"}:
        return False
    return default


def _load_driver_rows(uploaded) -> list[dict]:
    filename = (uploaded.name or "").lower()
    rows: list[dict] = []
    if filename.endswith(".csv"):
        uploaded.seek(0)
        content = uploaded.read().decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(content.splitlines())
        for row in reader:
            parsed = {}
            for key, value in (row or {}).items():
                if not key:
                    continue
                parsed[_map_driver_header(str(key))] = value
            rows.append(parsed)
        return rows
    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        uploaded.seek(0)
        wb = load_workbook(uploaded, read_only=True, data_only=True)
        ws = wb.active
        values = list(ws.values)
        if not values:
            return []
        headers = [_map_driver_header(str(h or "")) for h in values[0]]
        for raw in values[1:]:
            parsed = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                parsed[header] = raw[idx] if idx < len(raw) else None
            rows.append(parsed)
        return rows
    raise ValueError("Formato no soportado. Usa CSV o XLSX.")


@login_required
@permission_required("recetas.change_lineareceta", raise_exception=True)
def linea_edit(request: HttpRequest, pk: int, linea_id: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    linea = get_object_or_404(LineaReceta, pk=linea_id, receta=receta)

    if request.method == "POST":
        advanced_mode = receta.tipo == Receta.TIPO_PRODUCTO_FINAL and request.POST.get("advanced_mode") == "1"
        component_filter = (request.POST.get("component_filter") or "").strip().upper()
        component_context = (request.POST.get("component_context") or "").strip().lower()
        quick_mode = receta.tipo == Receta.TIPO_PRODUCTO_FINAL and not advanced_mode
        original_insumo_id = linea.insumo_id
        insumo_id = request.POST.get("insumo_id")
        selected_insumo = canonical_insumo_by_id(insumo_id)
        selected_was_variant = bool(selected_insumo and str(selected_insumo.id) != str(insumo_id or "").strip())
        tipo_linea = (request.POST.get("tipo_linea") or LineaReceta.TIPO_NORMAL).strip()
        if quick_mode:
            tipo_linea = LineaReceta.TIPO_NORMAL
        if tipo_linea not in {LineaReceta.TIPO_NORMAL, LineaReceta.TIPO_SUBSECCION}:
            tipo_linea = LineaReceta.TIPO_NORMAL
        linea.insumo_texto = (request.POST.get("insumo_texto") or "").strip()[:250]
        if not linea.insumo_texto and selected_insumo:
            linea.insumo_texto = _insumo_display_name(selected_insumo)[:250]
        linea.unidad_texto = (request.POST.get("unidad_texto") or "").strip()[:40]
        linea.etapa = (request.POST.get("etapa") or "").strip()[:120]
        linea.tipo_linea = tipo_linea
        linea.cantidad = _to_decimal_or_none(request.POST.get("cantidad"))
        linea.insumo = selected_insumo
        _autolink_pan_derived_from_recipe(receta, linea)
        canonicalized = _repoint_linea_to_canonical_if_needed(linea)
        if original_insumo_id != linea.insumo_id or canonicalized:
            # Si cambió el insumo ligado (manual/autolink), invalidamos snapshot previo
            # para recalcular costo unitario con el nuevo insumo.
            linea.costo_unitario_snapshot = None
            linea.costo_linea_excel = None
        if linea.insumo:
            linea.insumo_texto = _insumo_display_name(linea.insumo)[:250]

        if linea.insumo:
            linea.match_status = LineaReceta.STATUS_AUTO
            linea.match_method = "MANUAL"
            linea.match_score = 100.0
            linea.aprobado_por = request.user
            linea.aprobado_en = timezone.now()
        elif tipo_linea == LineaReceta.TIPO_SUBSECCION:
            linea.match_status = LineaReceta.STATUS_AUTO
            linea.match_method = LineaReceta.MATCH_SUBSECTION
            linea.match_score = 100.0
            linea.aprobado_por = request.user
            linea.aprobado_en = timezone.now()
        else:
            linea.match_status = LineaReceta.STATUS_REJECTED
            linea.match_method = LineaReceta.MATCH_NONE
            linea.match_score = 0.0

        _autofill_unidad_from_insumo(linea)
        validation_error = _validate_linea_operativa(receta, linea)
        validation_error = validation_error or _validate_component_filter_selection(receta, linea, component_filter)
        validation_error = validation_error or _validate_quick_mode_selection(receta, linea, quick_mode)
        validation_error = validation_error or _validate_selected_insumo_enterprise_ready(linea)
        if validation_error:
            messages.error(request, validation_error)
            return render(
                request,
                "recetas/linea_form.html",
                _linea_form_context(
                    receta,
                    linea,
                    advanced_mode=advanced_mode,
                    component_filter=component_filter,
                    component_context=component_context,
                ),
            )
        _switch_line_to_internal_cost(linea)
        linea.save()
        _sync_derived_insumos_safe(request, receta)
        _sync_cost_version_safe(request, receta, "LINEA_EDIT")
        if canonicalized or selected_was_variant:
            messages.info(request, f"El artículo se normalizó automáticamente a {_insumo_display_name(linea.insumo)}.")
        messages.success(request, "Línea actualizada.")
        return redirect("recetas:receta_detail", pk=pk)

    advanced_mode = receta.tipo == Receta.TIPO_PRODUCTO_FINAL and request.GET.get("advanced") == "1"
    component_filter = (request.GET.get("component_kind") or request.GET.get("component_filter") or "").strip().upper()
    component_context = (request.GET.get("component_context") or "").strip().lower()
    return render(
        request,
        "recetas/linea_form.html",
        _linea_form_context(
            receta,
            linea,
            advanced_mode=advanced_mode,
            component_filter=component_filter,
            component_context=component_context,
        ),
    )


@login_required
@permission_required("recetas.add_lineareceta", raise_exception=True)
def linea_create(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    if request.method == "POST":
        advanced_mode = receta.tipo == Receta.TIPO_PRODUCTO_FINAL and request.POST.get("advanced_mode") == "1"
        component_filter = (request.POST.get("component_filter") or "").strip().upper()
        component_context = (request.POST.get("component_context") or "").strip().lower()
        quick_mode = receta.tipo == Receta.TIPO_PRODUCTO_FINAL and not advanced_mode
        insumo_id = request.POST.get("insumo_id")
        selected_insumo = canonical_insumo_by_id(insumo_id)
        selected_was_variant = bool(selected_insumo and str(selected_insumo.id) != str(insumo_id or "").strip())
        tipo_linea = (request.POST.get("tipo_linea") or LineaReceta.TIPO_NORMAL).strip()
        if quick_mode:
            tipo_linea = LineaReceta.TIPO_NORMAL
        if tipo_linea not in {LineaReceta.TIPO_NORMAL, LineaReceta.TIPO_SUBSECCION}:
            tipo_linea = LineaReceta.TIPO_NORMAL
        posicion_default = (receta.lineas.order_by("-posicion").first().posicion + 1) if receta.lineas.exists() else 1
        insumo_texto = (request.POST.get("insumo_texto") or "").strip()[:250]
        if not insumo_texto and selected_insumo:
            insumo_texto = _insumo_display_name(selected_insumo)[:250]
        linea = LineaReceta(
            receta=receta,
            posicion=posicion_default,
            tipo_linea=tipo_linea,
            etapa=(request.POST.get("etapa") or "").strip()[:120],
            insumo_texto=insumo_texto,
            unidad_texto=(request.POST.get("unidad_texto") or "").strip()[:40],
            cantidad=_to_decimal_or_none(request.POST.get("cantidad")),
            costo_linea_excel=None,
            insumo=selected_insumo,
            unidad=None,
        )
        original_insumo_id = linea.insumo_id
        _autolink_pan_derived_from_recipe(receta, linea)
        canonicalized = _repoint_linea_to_canonical_if_needed(linea)
        if original_insumo_id != linea.insumo_id or canonicalized:
            linea.costo_unitario_snapshot = None
            linea.costo_linea_excel = None
        if linea.insumo:
            linea.insumo_texto = _insumo_display_name(linea.insumo)[:250]
        if linea.insumo:
            linea.match_status = LineaReceta.STATUS_AUTO
            linea.match_method = "MANUAL"
            linea.match_score = 100.0
            linea.aprobado_por = request.user
            linea.aprobado_en = timezone.now()
        elif tipo_linea == LineaReceta.TIPO_SUBSECCION:
            linea.match_status = LineaReceta.STATUS_AUTO
            linea.match_method = LineaReceta.MATCH_SUBSECTION
            linea.match_score = 100.0
            linea.aprobado_por = request.user
            linea.aprobado_en = timezone.now()
        else:
            linea.match_status = LineaReceta.STATUS_REJECTED
            linea.match_method = LineaReceta.MATCH_NONE
            linea.match_score = 0.0

        _autofill_unidad_from_insumo(linea)
        validation_error = _validate_linea_operativa(receta, linea)
        validation_error = validation_error or _validate_component_filter_selection(receta, linea, component_filter)
        validation_error = validation_error or _validate_quick_mode_selection(receta, linea, quick_mode)
        validation_error = validation_error or _validate_selected_insumo_enterprise_ready(linea)
        if validation_error:
            messages.error(request, validation_error)
            return render(
                request,
                "recetas/linea_form.html",
                _linea_form_context(
                    receta,
                    linea,
                    advanced_mode=advanced_mode,
                    component_filter=component_filter,
                    component_context=component_context,
                ),
            )
        _switch_line_to_internal_cost(linea)
        linea.save()
        _sync_derived_insumos_safe(request, receta)
        _sync_cost_version_safe(request, receta, "LINEA_CREATE")
        if canonicalized or selected_was_variant:
            messages.info(request, f"El artículo se normalizó automáticamente a {_insumo_display_name(linea.insumo)}.")
        messages.success(request, "Línea agregada.")
        return redirect("recetas:receta_detail", pk=pk)
    advanced_mode = receta.tipo == Receta.TIPO_PRODUCTO_FINAL and request.GET.get("advanced") == "1"
    component_filter = (request.GET.get("component_kind") or request.GET.get("component_filter") or "").strip().upper()
    component_context = (request.GET.get("component_context") or "").strip().lower()
    return render(
        request,
        "recetas/linea_form.html",
        _linea_form_context(
            receta,
            advanced_mode=advanced_mode,
            component_filter=component_filter,
            component_context=component_context,
        ),
    )


@login_required
@permission_required("recetas.delete_lineareceta", raise_exception=True)
@require_POST
def linea_delete(request: HttpRequest, pk: int, linea_id: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    linea = get_object_or_404(LineaReceta, pk=linea_id, receta=receta)
    linea.delete()
    _sync_derived_insumos_safe(request, receta)
    _sync_cost_version_safe(request, receta, "LINEA_DELETE")
    messages.success(request, "Línea eliminada.")
    return redirect("recetas:receta_detail", pk=pk)


@login_required
@permission_required("recetas.change_lineareceta", raise_exception=True)
@require_POST
def linea_apply_direct_base_replacement(request: HttpRequest, pk: int, linea_id: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    linea = get_object_or_404(LineaReceta.objects.select_related("insumo", "unidad"), pk=linea_id, receta=receta)
    linea.source_recipe = None
    linea.source_code_kind = None
    linea.source_active_presentaciones_count = 0
    linea.uses_direct_base_in_final = False
    if linea.insumo_id:
        linea.source_code_kind = _derived_code_kind(linea.insumo.codigo or "")
        source_recipe_id = _recipe_id_from_derived_code(linea.insumo.codigo or "")
        if source_recipe_id:
            linea.source_recipe = (
                Receta.objects.filter(pk=source_recipe_id).only("id", "nombre", "usa_presentaciones").first()
            )
            linea.source_active_presentaciones_count = RecetaPresentacion.objects.filter(
                receta_id=source_recipe_id,
                activo=True,
            ).count()
            linea.uses_direct_base_in_final = bool(
                receta.tipo == Receta.TIPO_PRODUCTO_FINAL
                and linea.insumo.tipo_item == Insumo.TIPO_INTERNO
                and linea.source_code_kind == "PREPARACION"
                and linea.source_recipe
                and linea.source_recipe.usa_presentaciones
                and linea.source_active_presentaciones_count > 0
            )

    replacement = _suggest_direct_base_replacement(linea)
    if not replacement:
        messages.warning(
            request,
            "No se encontró una presentación derivada sugerida para esta línea.",
        )
        return redirect("recetas:receta_detail", pk=pk)

    original_insumo = linea.insumo
    original_cantidad = Decimal(str(linea.cantidad or 0))
    _apply_direct_base_replacement_to_line(linea, replacement)
    linea.save(
        update_fields=[
            "insumo",
            "insumo_texto",
            "unidad",
            "unidad_texto",
            "cantidad",
            "costo_unitario_snapshot",
            "costo_linea_excel",
            "match_status",
            "match_method",
            "match_score",
        ]
    )
    _sync_cost_version_safe(request, receta, "LINEA_DIRECT_BASE_REPLACEMENT")
    log_event(
        request.user,
        "LINEA_DIRECT_BASE_REPLACEMENT",
        "recetas.LineaReceta",
        str(linea.id),
        {
            "receta_id": receta.id,
            "linea_id": linea.id,
            "from_insumo_id": original_insumo.id if original_insumo else None,
            "to_insumo_id": replacement["insumo"].id,
            "from_cantidad": str(original_cantidad),
            "to_cantidad": str(replacement["replacement_quantity"]),
            "presentacion_id": replacement["presentacion"].id,
        },
    )
    messages.success(
        request,
        f"Línea actualizada a {replacement['insumo'].nombre} ({replacement['replacement_quantity']} {linea.unidad_texto}).",
    )
    return redirect("recetas:receta_detail", pk=pk)


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
def presentacion_create(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    workflow_errors = _validate_presentacion_workflow(receta)
    if workflow_errors:
        for error in workflow_errors:
            messages.error(request, error)
        return redirect("recetas:receta_detail", pk=pk)
    if request.method == "POST":
        nombre = (request.POST.get("nombre") or "").strip()
        peso_por_unidad_kg = _to_decimal_or_none(request.POST.get("peso_por_unidad_kg"))
        activo = request.POST.get("activo") == "on"
        if not nombre:
            messages.error(request, "El nombre de la presentación es obligatorio.")
            return redirect("recetas:presentacion_create", pk=pk)
        if not peso_por_unidad_kg or peso_por_unidad_kg <= 0:
            messages.error(request, "Cantidad por presentación debe ser mayor que cero.")
            return redirect("recetas:presentacion_create", pk=pk)

        presentacion, _ = RecetaPresentacion.objects.update_or_create(
            receta=receta,
            nombre=nombre[:80],
            defaults={
                "peso_por_unidad_kg": peso_por_unidad_kg,
                "activo": activo,
            },
        )
        if not receta.usa_presentaciones:
            receta.usa_presentaciones = True
            receta.save(update_fields=["usa_presentaciones"])
        _sync_derived_insumos_safe(request, receta)
        _sync_cost_version_safe(request, receta, "PRESENTACION_CREATE")
        messages.success(request, "Presentación guardada.")
        return redirect("recetas:receta_detail", pk=pk)

    return render(
        request,
        "recetas/presentacion_form.html",
        {
            "receta": receta,
            "presentacion": None,
            "rendimiento_unit_code": (receta.rendimiento_unidad.codigo if receta.rendimiento_unidad else ""),
        },
    )


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
def presentacion_edit(request: HttpRequest, pk: int, presentacion_id: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    workflow_errors = _validate_presentacion_workflow(receta)
    if workflow_errors:
        for error in workflow_errors:
            messages.error(request, error)
        return redirect("recetas:receta_detail", pk=pk)
    presentacion = get_object_or_404(RecetaPresentacion, pk=presentacion_id, receta=receta)
    if request.method == "POST":
        nombre = (request.POST.get("nombre") or "").strip()
        peso_por_unidad_kg = _to_decimal_or_none(request.POST.get("peso_por_unidad_kg"))
        activo = request.POST.get("activo") == "on"
        if not nombre:
            messages.error(request, "El nombre de la presentación es obligatorio.")
            return redirect("recetas:presentacion_edit", pk=pk, presentacion_id=presentacion_id)
        if not peso_por_unidad_kg or peso_por_unidad_kg <= 0:
            messages.error(request, "Cantidad por presentación debe ser mayor que cero.")
            return redirect("recetas:presentacion_edit", pk=pk, presentacion_id=presentacion_id)

        presentacion.nombre = nombre[:80]
        presentacion.peso_por_unidad_kg = peso_por_unidad_kg
        presentacion.activo = activo
        presentacion.save()
        _sync_derived_insumos_safe(request, receta)
        _sync_cost_version_safe(request, receta, "PRESENTACION_EDIT")
        messages.success(request, "Presentación actualizada.")
        return redirect("recetas:receta_detail", pk=pk)

    return render(
        request,
        "recetas/presentacion_form.html",
        {
            "receta": receta,
            "presentacion": presentacion,
            "rendimiento_unit_code": (receta.rendimiento_unidad.codigo if receta.rendimiento_unidad else ""),
        },
    )


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def presentacion_delete(request: HttpRequest, pk: int, presentacion_id: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    workflow_errors = _validate_presentacion_workflow(receta)
    if workflow_errors:
        for error in workflow_errors:
            messages.error(request, error)
        return redirect("recetas:receta_detail", pk=pk)
    presentacion = get_object_or_404(RecetaPresentacion, pk=presentacion_id, receta=receta)
    try:
        sync_presentacion_insumo(presentacion, deactivate=True)
    except Exception:
        messages.warning(
            request,
            "La presentación se eliminó, pero falló la desactivación del insumo derivado.",
        )
    presentacion.delete()
    _sync_cost_version_safe(request, receta, "PRESENTACION_DELETE")
    messages.success(request, "Presentación eliminada.")
    return redirect("recetas:receta_detail", pk=pk)


@login_required
@permission_required("recetas.view_receta", raise_exception=True)
def receta_versiones_export(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    export_format = (request.GET.get("format") or "csv").strip().lower()
    try:
        versiones = _load_versiones_costeo(receta, 300)
    except (OperationalError, ProgrammingError):
        versiones = []

    if export_format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "versiones_costeo"
        ws.append(
            [
                "version",
                "fecha",
                "fuente",
                "driver_scope",
                "driver_nombre",
                "costo_mp",
                "costo_mo",
                "costo_indirecto",
                "costo_total",
                "rendimiento",
                "unidad_rendimiento",
                "costo_por_unidad_rendimiento",
            ]
        )
        for v in versiones:
            ws.append(
                [
                    v.version_num,
                    timezone.localtime(v.creado_en).strftime("%Y-%m-%d %H:%M"),
                    v.fuente,
                    v.driver_scope,
                    v.driver_nombre,
                    float(v.costo_mp or 0),
                    float(v.costo_mo or 0),
                    float(v.costo_indirecto or 0),
                    float(v.costo_total or 0),
                    float(v.rendimiento_cantidad or 0),
                    v.rendimiento_unidad,
                    float(v.costo_por_unidad_rendimiento or 0),
                ]
            )
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        resp = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="receta_{receta.id}_versiones_costeo.xlsx"'
        return resp

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="receta_{receta.id}_versiones_costeo.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "version",
            "fecha",
            "fuente",
            "driver_scope",
            "driver_nombre",
            "costo_mp",
            "costo_mo",
            "costo_indirecto",
            "costo_total",
            "rendimiento",
            "unidad_rendimiento",
            "costo_por_unidad_rendimiento",
        ]
    )
    for v in versiones:
        writer.writerow(
            [
                v.version_num,
                timezone.localtime(v.creado_en).strftime("%Y-%m-%d %H:%M"),
                v.fuente,
                v.driver_scope,
                v.driver_nombre,
                v.costo_mp,
                v.costo_mo,
                v.costo_indirecto,
                v.costo_total,
                v.rendimiento_cantidad,
                v.rendimiento_unidad,
                v.costo_por_unidad_rendimiento,
            ]
        )
    return response


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
def drivers_costeo(request: HttpRequest) -> HttpResponse:
    q = (request.GET.get("q") or "").strip()
    scope = (request.GET.get("scope") or "").strip().upper()
    edit_raw = (request.GET.get("edit") or "").strip()
    edit_driver = None

    drivers_unavailable = False
    drivers = []
    try:
        qs = CostoDriver.objects.select_related("receta").order_by("scope", "prioridad", "id")
        if q:
            qs = qs.filter(
                Q(nombre__icontains=q)
                | Q(familia__icontains=q)
                | Q(receta__nombre__icontains=q)
            )
        if scope in {CostoDriver.SCOPE_PRODUCTO, CostoDriver.SCOPE_FAMILIA, CostoDriver.SCOPE_LOTE, CostoDriver.SCOPE_GLOBAL}:
            qs = qs.filter(scope=scope)
        if edit_raw:
            try:
                edit_driver = CostoDriver.objects.select_related("receta").filter(pk=int(edit_raw)).first()
            except Exception:
                edit_driver = None
        drivers = list(qs[:400])
    except (OperationalError, ProgrammingError):
        drivers_unavailable = True
        edit_driver = None

    if request.method == "POST":
        if drivers_unavailable:
            messages.error(request, "Drivers de costeo no disponibles en este entorno. Ejecuta migraciones.")
            return redirect("recetas:drivers_costeo")
        driver_id = (request.POST.get("driver_id") or "").strip()
        nombre = (request.POST.get("nombre") or "").strip()
        driver_scope = _normalize_driver_scope(request.POST.get("scope"))
        receta_id = (request.POST.get("receta_id") or "").strip()
        familia = (request.POST.get("familia") or "").strip()
        lote_desde = _to_decimal_or_none(request.POST.get("lote_desde"))
        lote_hasta = _to_decimal_or_none(request.POST.get("lote_hasta"))
        mo_pct = _to_decimal_safe(request.POST.get("mo_pct"))
        indirecto_pct = _to_decimal_safe(request.POST.get("indirecto_pct"))
        mo_fijo = _to_decimal_safe(request.POST.get("mo_fijo"))
        indirecto_fijo = _to_decimal_safe(request.POST.get("indirecto_fijo"))
        prioridad = _to_int_safe(request.POST.get("prioridad"), 100)
        activo = _to_bool_safe(request.POST.get("activo"), True)

        if not nombre:
            messages.error(request, "Nombre del driver es obligatorio.")
            return redirect("recetas:drivers_costeo")

        receta = Receta.objects.filter(pk=receta_id).first() if receta_id else None
        if driver_scope == CostoDriver.SCOPE_PRODUCTO and not receta:
            messages.error(request, "Para scope PRODUCTO debes seleccionar una receta.")
            return redirect("recetas:drivers_costeo")

        try:
            if driver_id:
                driver = CostoDriver.objects.filter(pk=driver_id).first()
                if not driver:
                    messages.error(request, "Driver no encontrado para editar.")
                    return redirect("recetas:drivers_costeo")
            else:
                driver = CostoDriver()

            driver.nombre = nombre[:120]
            driver.scope = driver_scope
            driver.receta = receta
            driver.familia = familia[:120]
            driver.lote_desde = lote_desde
            driver.lote_hasta = lote_hasta
            driver.mo_pct = mo_pct
            driver.indirecto_pct = indirecto_pct
            driver.mo_fijo = mo_fijo
            driver.indirecto_fijo = indirecto_fijo
            driver.prioridad = max(prioridad, 0)
            driver.activo = activo
            driver.save()
            messages.success(request, "Driver guardado.")
        except (OperationalError, ProgrammingError):
            messages.error(request, "No se pudo guardar el driver. Ejecuta migraciones pendientes.")
        return redirect("recetas:drivers_costeo")

    return render(
        request,
        "recetas/drivers_costeo.html",
        {
            "drivers": drivers,
            "recetas": Receta.objects.order_by("nombre"),
            "q": q,
            "scope": scope,
            "edit_driver": edit_driver,
            "drivers_unavailable": drivers_unavailable,
            "scope_choices": [
                CostoDriver.SCOPE_PRODUCTO,
                CostoDriver.SCOPE_FAMILIA,
                CostoDriver.SCOPE_LOTE,
                CostoDriver.SCOPE_GLOBAL,
            ],
        },
    )


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def drivers_costeo_delete(request: HttpRequest, driver_id: int) -> HttpResponse:
    try:
        driver = get_object_or_404(CostoDriver, pk=driver_id)
        driver.delete()
        messages.success(request, "Driver eliminado.")
    except (OperationalError, ProgrammingError):
        messages.error(request, "No se pudo eliminar el driver. Ejecuta migraciones pendientes.")
    return redirect("recetas:drivers_costeo")


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
def drivers_costeo_plantilla(request: HttpRequest) -> HttpResponse:
    export_format = (request.GET.get("format") or "xlsx").strip().lower()
    headers = [
        "scope",
        "nombre",
        "receta",
        "codigo_point",
        "familia",
        "lote_desde",
        "lote_hasta",
        "mo_pct",
        "indirecto_pct",
        "mo_fijo",
        "indirecto_fijo",
        "prioridad",
        "activo",
    ]
    rows = [
        ["PRODUCTO", "Driver Pastel Fresas", "Pastel Fresas Con Crema - Chico", "PFC-CHICO", "", "", "", "8", "4", "0", "0", "10", "1"],
        ["FAMILIA", "Driver Insumos 1", "", "", "Insumos 1", "", "", "6", "3", "0", "0", "30", "1"],
        ["GLOBAL", "Driver Global Base", "", "", "", "", "", "5", "2", "0", "0", "100", "1"],
    ]

    if export_format == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="plantilla_drivers_costeo.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response

    wb = Workbook()
    ws = wb.active
    ws.title = "drivers_costeo"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="F8E8EE")
    header_font = Font(color="8B2252", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.auto_filter.ref = f"A1:M{len(rows) + 1}"
    ws.freeze_panes = "A2"

    widths = {
        "A": 14,   # scope
        "B": 30,   # nombre
        "C": 32,   # receta
        "D": 20,   # codigo_point
        "E": 20,   # familia
        "F": 12,   # lote_desde
        "G": 12,   # lote_hasta
        "H": 10,   # mo_pct
        "I": 12,   # indirecto_pct
        "J": 10,   # mo_fijo
        "K": 14,   # indirecto_fijo
        "L": 10,   # prioridad
        "M": 8,    # activo
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws_guide = wb.create_sheet("instrucciones")
    ws_guide.append(["Campo", "Obligatorio", "Descripción / valores válidos"])
    guide_rows = [
        ["scope", "Sí", "GLOBAL | FAMILIA | PRODUCTO | LOTE"],
        ["nombre", "Sí", "Nombre descriptivo del driver."],
        ["receta", "Condicional", "Obligatorio si scope=PRODUCTO."],
        ["codigo_point", "Opcional", "Alias de receta para encontrarla por código comercial."],
        ["familia", "Condicional", "Obligatorio si scope=FAMILIA."],
        ["lote_desde", "Condicional", "Usar con scope=LOTE (decimal >= 0)."],
        ["lote_hasta", "Condicional", "Usar con scope=LOTE (decimal >= lote_desde)."],
        ["mo_pct", "Sí", "Porcentaje de mano de obra, por ejemplo 8."],
        ["indirecto_pct", "Sí", "Porcentaje de indirectos, por ejemplo 4."],
        ["mo_fijo", "Opcional", "Monto fijo adicional de MO."],
        ["indirecto_fijo", "Opcional", "Monto fijo adicional de indirectos."],
        ["prioridad", "Sí", "Menor número = mayor prioridad."],
        ["activo", "Sí", "1 activo / 0 inactivo."],
    ]
    for row in guide_rows:
        ws_guide.append(row)
    for col in range(1, 4):
        cell = ws_guide.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    ws_guide.column_dimensions["A"].width = 22
    ws_guide.column_dimensions["B"].width = 14
    ws_guide.column_dimensions["C"].width = 72
    ws_guide.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_drivers_costeo.xlsx"'
    return response


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def drivers_costeo_importar(request: HttpRequest) -> HttpResponse:
    try:
        CostoDriver.objects.exists()
    except (OperationalError, ProgrammingError):
        messages.error(request, "Drivers de costeo no disponibles en este entorno. Ejecuta migraciones.")
        return redirect("recetas:drivers_costeo")

    uploaded = request.FILES.get("archivo")
    modo = (request.POST.get("modo") or "upsert").strip().lower()
    if modo not in {"upsert", "replace"}:
        modo = "upsert"

    if not uploaded:
        messages.error(request, "Selecciona un archivo para importar drivers.")
        return redirect("recetas:drivers_costeo")

    try:
        rows = _load_driver_rows(uploaded)
    except Exception as exc:
        messages.error(request, f"No se pudo leer archivo de drivers: {exc}")
        return redirect("recetas:drivers_costeo")

    if not rows:
        messages.warning(request, "Archivo de drivers sin filas.")
        return redirect("recetas:drivers_costeo")

    created = 0
    updated = 0
    skipped = 0

    if modo == "replace":
        CostoDriver.objects.all().delete()

    for row in rows:
        scope = _normalize_driver_scope(row.get("scope"))
        nombre = str(row.get("nombre") or "").strip()
        if not nombre:
            skipped += 1
            continue

        receta = None
        receta_name = str(row.get("receta") or "").strip()
        codigo_point = str(row.get("codigo_point") or "").strip()
        if codigo_point:
            receta = Receta.objects.filter(codigo_point__iexact=codigo_point).order_by("id").first()
        if receta is None and receta_name:
            receta = Receta.objects.filter(nombre_normalizado=normalizar_nombre(receta_name)).order_by("id").first()

        if scope == CostoDriver.SCOPE_PRODUCTO and not receta:
            skipped += 1
            continue

        familia = str(row.get("familia") or "").strip()
        lote_desde = _to_decimal_or_none(str(row.get("lote_desde") or "").strip())
        lote_hasta = _to_decimal_or_none(str(row.get("lote_hasta") or "").strip())

        filter_kwargs = {
            "scope": scope,
            "nombre": nombre[:120],
            "receta": receta,
            "familia_normalizada": normalizar_nombre(familia),
            "lote_desde": lote_desde,
            "lote_hasta": lote_hasta,
        }
        driver = CostoDriver.objects.filter(**filter_kwargs).first()
        if not driver:
            driver = CostoDriver(**filter_kwargs)
            created += 1
        else:
            updated += 1

        driver.familia = familia[:120]
        driver.mo_pct = _to_decimal_safe(row.get("mo_pct"))
        driver.indirecto_pct = _to_decimal_safe(row.get("indirecto_pct"))
        driver.mo_fijo = _to_decimal_safe(row.get("mo_fijo"))
        driver.indirecto_fijo = _to_decimal_safe(row.get("indirecto_fijo"))
        driver.prioridad = max(_to_int_safe(row.get("prioridad"), 100), 0)
        driver.activo = _to_bool_safe(row.get("activo"), True)
        driver.save()

    messages.success(
        request,
        f"Importación drivers completada. Creados: {created}. Actualizados: {updated}. Omitidos: {skipped}.",
    )
    return redirect("recetas:drivers_costeo")


@login_required
@permission_required("recetas.change_lineareceta", raise_exception=True)
def matching_pendientes(request: HttpRequest) -> HttpResponse:
    q = request.GET.get("q", "").strip()
    receta_id_raw = (request.GET.get("receta") or "").strip()
    receta_filter = None
    if receta_id_raw.isdigit():
        receta_filter = int(receta_id_raw)
    pendientes = LineaReceta.objects.filter(match_status=LineaReceta.STATUS_NEEDS_REVIEW).select_related("receta", "insumo").order_by("receta__nombre", "posicion")
    if q:
        pendientes = pendientes.filter(insumo_texto__icontains=q)
    if receta_filter:
        pendientes = pendientes.filter(receta_id=receta_filter)

    total_pendientes = pendientes.count()
    recetas_afectadas = pendientes.values("receta_id").distinct().count()
    no_match_count = pendientes.filter(match_method=LineaReceta.MATCH_NONE).count()
    fuzzy_count = pendientes.filter(match_method=LineaReceta.MATCH_FUZZY).count()
    governance_rows = [
        {
            "front": "Detección BOM",
            "owner": "Producción / Recetas",
            "blockers": total_pendientes,
            "completion": 100 if total_pendientes == 0 else 35,
            "detail": (
                "No hay partidas abiertas en revisión."
                if total_pendientes == 0
                else f"{total_pendientes} partida(s) siguen sin cierre sobre artículo maestro."
            ),
            "next_step": (
                "Mantener control preventivo del BOM."
                if total_pendientes == 0
                else "Revisar partidas abiertas y confirmar sugerencias."
            ),
            "url": reverse("recetas:matching_pendientes"),
            "cta": "Abrir bandeja",
        },
        {
            "front": "Artículo maestro",
            "owner": "Maestros / Inventario",
            "blockers": no_match_count,
            "completion": 100 if no_match_count == 0 else 70,
            "detail": (
                "Todas las partidas ya tienen artículo maestro disponible."
                if no_match_count == 0
                else f"{no_match_count} partida(s) siguen sin alta maestra."
            ),
            "next_step": (
                "Mantener catálogo estable."
                if no_match_count == 0
                else "Dar de alta o asignar artículo maestro."
            ),
            "url": reverse("maestros:insumo_list"),
            "cta": "Abrir maestro",
        },
        {
            "front": "Costeo y MRP",
            "owner": "ERP / Operaciones",
            "blockers": fuzzy_count,
            "completion": 100 if fuzzy_count == 0 else 85,
            "detail": (
                "Costeo y MRP pueden operar sin referencias abiertas."
                if fuzzy_count == 0
                else f"{fuzzy_count} partida(s) todavía dependen de validación operativa."
            ),
            "next_step": (
                "Monitorear operación documental."
                if fuzzy_count == 0
                else "Aplicar o confirmar la sugerencia operativa."
            ),
            "url": reverse("recetas:mrp_form"),
            "cta": "Abrir MRP",
        },
    ]
    workflow_rows = [
        {
            "step": "01",
            "title": "Detección BOM",
            "owner": "Producción / Recetas",
            "open": total_pendientes,
            "closed": 0 if total_pendientes else recetas_afectadas,
            "completion": 100 if total_pendientes == 0 else 35,
            "tone": "success" if total_pendientes == 0 else "warning",
            "detail": (
                "No hay componentes abiertos en revisión."
                if total_pendientes == 0
                else f"{total_pendientes} componente(s) siguen abiertos dentro del BOM."
            ),
            "next_step": (
                "Mantener control preventivo del BOM."
                if total_pendientes == 0
                else "Revisar componentes abiertos y confirmar sugerencias."
            ),
            "action_label": "Abrir bandeja",
            "action_href": reverse("recetas:matching_pendientes"),
        },
        {
            "step": "02",
            "title": "Artículo maestro",
            "owner": "Maestros / Inventario",
            "open": no_match_count,
            "closed": max(total_pendientes - no_match_count, 0),
            "completion": 100 if no_match_count == 0 else 70,
            "tone": "success" if no_match_count == 0 else "warning",
            "detail": (
                "Todos los componentes ya tienen artículo maestro disponible."
                if no_match_count == 0
                else f"{no_match_count} componente(s) siguen sin alta maestra."
            ),
            "next_step": (
                "Mantener catálogo maestro estable."
                if no_match_count == 0
                else "Dar de alta o asignar artículo maestro."
            ),
            "action_label": "Abrir maestro",
            "action_href": reverse("maestros:insumo_list"),
        },
        {
            "step": "03",
            "title": "Cierre operativo",
            "owner": "ERP / Operaciones",
            "open": fuzzy_count,
            "closed": max(total_pendientes - fuzzy_count, 0),
            "completion": 100 if fuzzy_count == 0 else 85,
            "tone": "success" if fuzzy_count == 0 else "warning",
            "detail": (
                "Costeo y planeación ya operan sin referencias abiertas."
                if fuzzy_count == 0
                else f"{fuzzy_count} componente(s) todavía requieren confirmación operativa."
            ),
            "next_step": (
                "Monitorear operación documental."
                if fuzzy_count == 0
                else "Aplicar sugerencias y cerrar la bandeja."
            ),
            "action_label": "Abrir MRP",
            "action_href": reverse("recetas:mrp_form"),
        },
    ]
    if total_pendientes == 0:
        erp_command_center = {
            "owner": "Producción / Maestros",
            "status": "Estable",
            "tone": "success",
            "blockers": 0,
            "next_step": "Mantener el control preventivo de componentes y validar nuevas altas maestras antes del costeo.",
            "url": reverse("recetas:recetas_list"),
            "cta": "Abrir catálogo de recetas",
        }
    elif no_match_count > 0:
        erp_command_center = {
            "owner": "Maestros / Inventario",
            "status": "Crítico",
            "tone": "danger",
            "blockers": total_pendientes,
            "next_step": "Dar de alta artículos maestros faltantes y cerrar primero los componentes sin alta.",
            "url": reverse("maestros:insumo_list"),
            "cta": "Abrir maestro de artículos",
        }
    else:
        erp_command_center = {
            "owner": "Producción / Operaciones",
            "status": "En revisión",
            "tone": "warning",
            "blockers": total_pendientes,
            "next_step": "Confirmar sugerencias operativas y cerrar las referencias abiertas para liberar costeo y planeación.",
            "url": reverse("recetas:matching_pendientes"),
            "cta": "Cerrar componentes",
        }

    export_format = (request.GET.get("export") or "").strip().lower()
    if export_format == "csv":
        now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="matching_pendientes_{now_str}.csv"'
        writer = csv.writer(response)
        writer.writerow(["receta", "posicion", "ingrediente", "metodo", "score", "insumo_ligado"])
        for linea in pendientes:
            writer.writerow(
                [
                    linea.receta.nombre,
                    linea.posicion,
                    linea.insumo_texto or "",
                    linea.match_method or "",
                    float(linea.match_score or 0),
                    linea.insumo.nombre if linea.insumo_id and linea.insumo else "",
                ]
            )
        return response
    if export_format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "matching_pendientes"
        ws.append(["receta", "posicion", "ingrediente", "metodo", "score", "insumo_ligado"])
        for linea in pendientes.iterator(chunk_size=500):
            ws.append(
                [
                    linea.receta.nombre,
                    linea.posicion,
                    linea.insumo_texto or "",
                    linea.match_method or "",
                    float(linea.match_score or 0),
                    linea.insumo.nombre if linea.insumo_id and linea.insumo else "",
                ]
            )
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 34
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 34

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="matching_pendientes_{now_str}.xlsx"'
        return response

    paginator = Paginator(pendientes, 25)
    page = paginator.get_page(request.GET.get("page"))
    auto_suggested_count = 0
    canonical_suggested_count = 0
    for linea in page.object_list:
        _attach_linea_suggested_match(linea)
        if linea.suggested_can_approve:
            auto_suggested_count += 1
        if linea.suggested_is_canonical:
            canonical_suggested_count += 1
    return render(
        request,
        "recetas/matching_pendientes.html",
        {
            "page": page,
            "q": q,
            "receta_filter": receta_filter,
            "erp_command_center": erp_command_center,
            "workflow_rows": workflow_rows,
            "erp_governance_rows": governance_rows,
            "executive_radar_rows": _recipes_executive_radar_rows(
                governance_rows,
                owner="Producción / Maestros",
                fallback_url=reverse("recetas:matching_pendientes"),
            ),
            "critical_path_rows": _recipes_critical_path_rows(
                governance_rows,
                owner="Producción / Maestros",
                fallback_url=reverse("recetas:matching_pendientes"),
            ),
            "stats": {
                "total": total_pendientes,
                "recetas": recetas_afectadas,
                "no_match": no_match_count,
                "fuzzy": fuzzy_count,
                "auto_suggested": auto_suggested_count,
                "canonical_suggested": canonical_suggested_count,
            },
        },
    )


@login_required
@permission_required("recetas.change_lineareceta", raise_exception=True)
@require_POST
def aprobar_matching_sugerido_lote(request: HttpRequest) -> HttpResponse:
    q = request.POST.get("q", "").strip()
    page_number = request.POST.get("page", "").strip()
    receta_id_raw = (request.POST.get("receta") or "").strip()
    receta_filter = int(receta_id_raw) if receta_id_raw.isdigit() else None

    pendientes = (
        LineaReceta.objects.filter(match_status=LineaReceta.STATUS_NEEDS_REVIEW)
        .select_related("receta", "insumo")
        .order_by("receta__nombre", "posicion")
    )
    if q:
        pendientes = pendientes.filter(insumo_texto__icontains=q)
    if receta_filter:
        pendientes = pendientes.filter(receta_id=receta_filter)

    paginator = Paginator(pendientes, 25)
    page = paginator.get_page(page_number or 1)

    aprobadas = 0
    for linea in page.object_list:
        suggested_insumo, raw_score, raw_method = match_insumo(linea.insumo_texto or "")
        canonical_insumo = _canonicalize_insumo_match(suggested_insumo)
        if not canonical_insumo or float(raw_score or 0.0) < 75.0:
            continue

        linea.insumo = canonical_insumo
        linea.match_status = LineaReceta.STATUS_AUTO
        linea.match_method = f"{raw_method}_CANON" if canonical_insumo != suggested_insumo else raw_method
        linea.match_score = 100.0 if canonical_insumo != suggested_insumo else float(raw_score or 0.0)
        linea.aprobado_por = request.user
        linea.aprobado_en = timezone.now()
        linea.save()
        _sync_cost_version_safe(request, linea.receta, "MATCHING_APPROVE_SUGGESTED_BULK")
        aprobadas += 1

    if aprobadas:
        messages.success(request, f"Depuración aplicada en lote: {aprobadas} línea(s) de la página actual.")
    else:
        messages.warning(request, "No había sugerencias aprobables en la página actual.")

    query = {}
    if q:
        query["q"] = q
    if receta_filter:
        query["receta"] = receta_filter
    if page_number:
        query["page"] = page_number
    url = reverse("recetas:matching_pendientes")
    if query:
        url = f"{url}?{urlencode(query)}"
    return redirect(url)


@login_required
@permission_required("recetas.change_lineareceta", raise_exception=True)
@require_POST
def receta_aprobar_sugeridos(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    aprobadas = 0
    pendientes = (
        receta.lineas.filter(match_status__in=[LineaReceta.STATUS_NEEDS_REVIEW, LineaReceta.STATUS_REJECTED])
        .select_related("insumo")
        .order_by("posicion")
    )
    for linea in pendientes:
        suggested_insumo, raw_score, raw_method = match_insumo(linea.insumo_texto or "")
        canonical_insumo = _canonicalize_insumo_match(suggested_insumo)
        if not canonical_insumo or float(raw_score or 0.0) < 75.0:
            continue

        linea.insumo = canonical_insumo
        linea.match_status = LineaReceta.STATUS_AUTO
        linea.match_method = f"{raw_method}_CANON" if canonical_insumo != suggested_insumo else raw_method
        linea.match_score = 100.0 if canonical_insumo != suggested_insumo else float(raw_score or 0.0)
        linea.aprobado_por = request.user
        linea.aprobado_en = timezone.now()
        linea.save()
        aprobadas += 1

    if aprobadas:
        _sync_cost_version_safe(request, receta, "MATCHING_APPROVE_SUGGESTED_RECIPE")
        messages.success(request, f"Receta actualizada: {aprobadas} línea(s) aprobadas con sugerencia maestra.")
    else:
        messages.warning(request, "La receta no tiene sugerencias aprobables en este momento.")
    return redirect("recetas:receta_detail", pk=receta.id)


@login_required
@permission_required("recetas.change_lineareceta", raise_exception=True)
@require_POST
def receta_repoint_canonical(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    actualizadas = 0
    for linea in receta.lineas.select_related("insumo").all().order_by("posicion"):
        if not linea.insumo_id:
            continue
        canonical = _canonicalize_insumo_match(linea.insumo)
        if not canonical or canonical.id == linea.insumo_id:
            continue
        linea.insumo = canonical
        linea.insumo_texto = canonical.nombre[:250]
        _autofill_unidad_from_insumo(linea)
        _switch_line_to_internal_cost(linea)
        linea.save()
        actualizadas += 1

    if actualizadas:
        _sync_cost_version_safe(request, receta, "RECIPE_REPOINT_CANONICAL")
        log_event(
            request.user,
            "RECIPE_REPOINT_CANONICAL",
            "Receta",
            str(receta.id),
            {"receta_id": receta.id, "updated_lines": actualizadas},
        )
        messages.success(request, f"Receta actualizada: {actualizadas} componente(s) apuntan ya al artículo estándar.")
    else:
        messages.info(request, "La receta ya estaba usando artículos maestros.")
    return redirect("recetas:receta_detail", pk=receta.id)


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def receta_apply_direct_base_replacements(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    lineas = list(receta.lineas.select_related("insumo", "unidad").all())
    applied = 0
    replacement_cache: dict[int, list[dict[str, object]]] = {}
    for linea in lineas:
        linea.source_recipe = None
        linea.source_code_kind = None
        linea.source_active_presentaciones_count = 0
        linea.uses_direct_base_in_final = False
        if linea.insumo_id:
            linea.source_code_kind = _derived_code_kind(linea.insumo.codigo or "")
            source_recipe_id = _recipe_id_from_derived_code(linea.insumo.codigo or "")
            if source_recipe_id:
                linea.source_recipe = (
                    Receta.objects.filter(pk=source_recipe_id).only("id", "nombre", "usa_presentaciones").first()
                )
                linea.source_active_presentaciones_count = RecetaPresentacion.objects.filter(
                    receta_id=source_recipe_id,
                    activo=True,
                ).count()
                linea.uses_direct_base_in_final = bool(
                    receta.tipo == Receta.TIPO_PRODUCTO_FINAL
                    and linea.insumo.tipo_item == Insumo.TIPO_INTERNO
                    and linea.source_code_kind == "PREPARACION"
                    and linea.source_recipe
                    and linea.source_recipe.usa_presentaciones
                    and linea.source_active_presentaciones_count > 0
                )
        replacement = _suggest_direct_base_replacement(linea, cache=replacement_cache)
        if not replacement:
            continue
        original_insumo = linea.insumo
        original_cantidad = Decimal(str(linea.cantidad or 0))
        _apply_direct_base_replacement_to_line(linea, replacement)
        linea.save(
            update_fields=[
                "insumo",
                "insumo_texto",
                "unidad",
                "unidad_texto",
                "cantidad",
                "costo_unitario_snapshot",
                "costo_linea_excel",
                "match_status",
                "match_method",
                "match_score",
            ]
        )
        log_event(
            request.user,
            "LINEA_DIRECT_BASE_REPLACEMENT",
            "recetas.LineaReceta",
            str(linea.id),
            {
                "receta_id": receta.id,
                "linea_id": linea.id,
                "from_insumo_id": original_insumo.id if original_insumo else None,
                "to_insumo_id": replacement["insumo"].id,
                "from_cantidad": str(original_cantidad),
                "to_cantidad": str(replacement["replacement_quantity"]),
                "presentacion_id": replacement["presentacion"].id,
                "batch_recipe_action": True,
            },
        )
        applied += 1

    if applied:
        _sync_cost_version_safe(request, receta, "RECIPE_DIRECT_BASE_REPLACEMENT")
        messages.success(request, f"Se actualizaron {applied} línea(s) a su presentación derivada sugerida.")
    else:
        messages.info(request, "No hubo líneas con sugerencia aplicable de derivado.")
    return redirect("recetas:receta_detail", pk=receta.id)


@login_required
@permission_required("recetas.change_lineareceta", raise_exception=True)
def matching_insumos_search(request: HttpRequest) -> JsonResponse:
    q = (request.GET.get("q") or "").strip()
    limit_raw = (request.GET.get("limit") or "20").strip()
    try:
        limit = int(limit_raw)
    except ValueError:
        limit = 20
    limit = max(1, min(limit, 100))

    queryset = Insumo.objects.filter(activo=True).select_related("unidad_base")
    if q:
        queryset = queryset.filter(nombre__icontains=q)
    grouped = {}
    for insumo in queryset.order_by("nombre")[: limit * 5]:
        key = insumo.nombre_normalizado or normalizar_nombre(insumo.nombre or "")
        grouped.setdefault(key, []).append(insumo)

    items = []
    for _, variants in grouped.items():
        canonical = _canonicalize_insumo_match(variants[0]) if variants else None
        if not canonical:
            continue
        items.append(
            {
                "id": canonical.id,
                "nombre": canonical.nombre,
                "label": (
                    f"{canonical.nombre} · Maestro ({len(variants)} variante(s))"
                    if len(variants) > 1
                    else canonical.nombre
                ),
                "variant_count": len(variants),
                "is_canonical": True,
            }
        )
        if len(items) >= limit:
            break
    return JsonResponse({"results": items, "count": len(items)})

@login_required
@permission_required("recetas.change_lineareceta", raise_exception=True)
def aprobar_matching(request: HttpRequest, linea_id: int) -> HttpResponse:
    linea = get_object_or_404(LineaReceta, pk=linea_id)
    next_url = (request.POST.get("next") or "").strip() or reverse("recetas:matching_pendientes")
    insumo_id = request.POST.get("insumo_id")
    if not insumo_id:
        messages.error(request, "Selecciona un insumo para aprobar.")
        return redirect(next_url)

    selected_raw = Insumo.objects.filter(pk=insumo_id).first()
    insumo = canonical_insumo_by_id(insumo_id)
    if not insumo:
        messages.error(request, "Selecciona un artículo válido para aprobar.")
        return redirect(next_url)
    canonical_insumo = _canonicalize_insumo_match(insumo)
    if not canonical_insumo:
        messages.error(request, "No se pudo determinar un artículo estándar para este componente.")
        return redirect(next_url)
    linea.insumo = canonical_insumo
    linea.match_status = LineaReceta.STATUS_AUTO
    raw_selected_id = selected_raw.id if selected_raw else None
    linea.match_method = "MANUAL_CANON" if raw_selected_id and canonical_insumo.id != raw_selected_id else "MANUAL"
    linea.match_score = 100.0
    linea.aprobado_por = request.user
    linea.aprobado_en = timezone.now()
    linea.save()
    _sync_cost_version_safe(request, linea.receta, "MATCHING_APPROVE")
    messages.success(request, f"Artículo asignado: {linea.insumo_texto} → {_insumo_display_name(canonical_insumo)}")
    return redirect(next_url)


@login_required
@permission_required("recetas.change_lineareceta", raise_exception=True)
@require_POST
def linea_repoint_canonical(request: HttpRequest, linea_id: int) -> HttpResponse:
    linea = get_object_or_404(LineaReceta.objects.select_related("receta", "insumo"), pk=linea_id)
    next_url = (request.POST.get("next") or "").strip() or reverse("recetas:receta_detail", args=[linea.receta.id])
    if not linea.insumo_id:
        messages.warning(request, "La línea no tiene artículo ligado para normalizar.")
        return redirect(next_url)

    canonical = _canonicalize_insumo_match(linea.insumo)
    if not canonical or canonical.id == linea.insumo_id:
        messages.info(request, "El componente ya está ligado al artículo estándar.")
        return redirect(next_url)

    previous_name = _insumo_display_name(linea.insumo)
    linea.insumo = canonical
    linea.insumo_texto = _insumo_display_name(canonical)[:250]
    _autofill_unidad_from_insumo(linea)
    _switch_line_to_internal_cost(linea)
    linea.save()
    _sync_cost_version_safe(request, linea.receta, "LINEA_REPOINT_CANONICAL")
    log_event(
        request.user,
        "LINEA_REPOINT_CANONICAL",
        "LineaReceta",
        str(linea.id),
        {
            "linea_id": linea.id,
            "receta_id": linea.receta.id,
            "from": previous_name,
            "to": _insumo_display_name(canonical),
        },
    )
    messages.success(request, f"Línea actualizada: {previous_name} → {_insumo_display_name(canonical)}")
    return redirect(next_url)


@login_required
@permission_required("recetas.change_lineareceta", raise_exception=True)
@require_POST
def aprobar_matching_sugerido(request: HttpRequest, linea_id: int) -> HttpResponse:
    linea = get_object_or_404(LineaReceta, pk=linea_id)
    next_url = (request.POST.get("next") or "").strip() or reverse("recetas:matching_pendientes")
    suggested_insumo, raw_score, raw_method = match_insumo(linea.insumo_texto or "")
    canonical_insumo = _canonicalize_insumo_match(suggested_insumo)

    if not canonical_insumo or float(raw_score or 0.0) < 75.0:
        messages.error(request, "Esta línea no tiene una sugerencia automática suficientemente confiable.")
        return redirect(next_url)

    linea.insumo = canonical_insumo
    linea.match_status = LineaReceta.STATUS_AUTO
    linea.match_method = f"{raw_method}_CANON" if canonical_insumo != suggested_insumo else raw_method
    linea.match_score = 100.0 if canonical_insumo != suggested_insumo else float(raw_score or 0.0)
    linea.aprobado_por = request.user
    linea.aprobado_en = timezone.now()
    linea.save()
    _sync_cost_version_safe(request, linea.receta, "MATCHING_APPROVE_SUGGESTED")
    messages.success(request, f"Sugerencia aplicada: {linea.insumo_texto} → {canonical_insumo.nombre}")
    return redirect(next_url)


def _linea_unit_code(linea: LineaReceta) -> str:
    if linea.unidad_id and linea.unidad:
        return linea.unidad.codigo
    txt = (linea.unidad_texto or "").strip()
    if txt:
        return txt
    if linea.insumo_id and linea.insumo and linea.insumo.unidad_base_id and linea.insumo.unidad_base:
        return linea.insumo.unidad_base.codigo
    return "-"


def _linea_unit_cost(linea: LineaReceta, unit_cost_cache: Dict[int, Decimal | None]) -> Decimal | None:
    if not linea.insumo_id:
        return None
    if linea.costo_unitario_snapshot is not None and linea.costo_unitario_snapshot > 0:
        return Decimal(str(linea.costo_unitario_snapshot))

    if linea.insumo_id in unit_cost_cache:
        return unit_cost_cache[linea.insumo_id]

    latest = _latest_cost_for_insumo(linea.insumo)
    unit_cost_cache[linea.insumo_id] = latest
    return latest


def _build_derived_parent_requirement(receta: Receta, multiplicador: Decimal) -> Dict[str, Any] | None:
    relation = get_active_derived_relation(receta)
    if relation is None or multiplicador <= 0:
        return None

    units_per_parent = Decimal(str(relation.unidades_por_padre or 0))
    if units_per_parent <= 0:
        return None

    parent_qty = multiplicador / units_per_parent
    if parent_qty <= 0:
        return None

    parent_recipe = relation.receta_padre
    parent_unit_cost = parent_recipe.costo_total_estimado_decimal
    inventory = InventarioCedisProducto.objects.filter(receta=parent_recipe).first()
    stock_actual = inventory.disponible if inventory else Decimal("0")
    faltante = parent_qty - Decimal(str(stock_actual or 0))
    faltante = faltante if faltante > 0 else Decimal("0")
    costo_total = parent_qty * parent_unit_cost if parent_unit_cost > 0 else Decimal("0")

    workflow_health_label = "Cubierto"
    workflow_health_tone = "success"
    workflow_action_label = "Abrir receta padre"
    workflow_action_url = reverse("recetas:receta_detail", args=[parent_recipe.id])
    workflow_action_method = "get"
    workflow_next = "El stock del producto padre cubre la presentación derivada."
    if parent_unit_cost <= 0:
        workflow_health_label = "Sin costo base padre"
        workflow_health_tone = "warning"
        workflow_next = "Cierra el costeo del producto padre antes de usar esta presentación en producción."
    elif faltante > 0:
        workflow_health_label = "Preparar padre"
        workflow_health_tone = "warning"
        workflow_next = "Programa producción o disponibilidad del producto padre para cubrir las rebanadas requeridas."

    return {
        "key": f"DERIVED_PARENT:{parent_recipe.id}",
        "row_kind": "DERIVED_PARENT",
        "is_derived_parent": True,
        "insumo": None,
        "insumo_id": None,
        "parent_recipe": parent_recipe,
        "parent_recipe_id": parent_recipe.id,
        "parent_recipe_name": parent_recipe.nombre,
        "nombre": f"{parent_recipe.nombre} (producto padre prorrateado)",
        "origen": "Interno",
        "display_origen": "Producto padre",
        "article_class_key": "DERIVED_PARENT",
        "article_class_label": "Producto padre",
        "proveedor_sugerido": "-",
        "unidad": "pza",
        "cantidad": parent_qty,
        "costo_total": costo_total,
        "costo": costo_total,
        "costo_unitario": parent_unit_cost,
        "stock_actual": Decimal(str(stock_actual or 0)),
        "faltante": faltante,
        "alerta_capacidad": faltante > 0,
        "master_missing": [],
        "workflow_health_label": workflow_health_label,
        "workflow_health_tone": workflow_health_tone,
        "workflow_action_label": workflow_action_label,
        "workflow_action_url": workflow_action_url,
        "workflow_action_method": workflow_action_method,
        "workflow_next": workflow_next,
        "source_recipe_name": parent_recipe.nombre,
        "source_recipe_id": parent_recipe.id,
        "derived_units_per_parent": units_per_parent,
        "detail_url": workflow_action_url,
    }


def _apply_producto_cedis_movimiento(movimiento: MovimientoProductoCedis) -> None:
    inventario, _ = InventarioCedisProducto.objects.get_or_create(receta=movimiento.receta)
    if movimiento.tipo == MovimientoProductoCedis.TIPO_ENTRADA:
        inventario.stock_actual += movimiento.cantidad
    else:
        inventario.stock_actual -= movimiento.cantidad
    inventario.actualizado_en = timezone.now()
    inventario.save(update_fields=["stock_actual", "actualizado_en"])


def _plan_consumo_can_apply(explosion: Dict[str, Any]) -> tuple[bool, str]:
    if int(explosion.get("lineas_sin_match") or 0) > 0:
        return False, "El plan tiene componentes sin artículo estándar."
    if len(explosion.get("lineas_sin_cantidad") or []) > 0:
        return False, "El plan tiene cantidades pendientes en su BOM."
    if len(explosion.get("lineas_sin_costo_unitario") or []) > 0:
        return False, "El plan tiene costos pendientes de definir."
    if int(explosion.get("alertas_capacidad") or 0) > 0:
        return False, "El plan no tiene stock suficiente para aplicar consumo real."
    return True, ""


def _mark_plan_consumption_applied(plan: PlanProduccion, acted_by) -> bool:
    update_fields: list[str] = []
    if plan.estado != PlanProduccion.ESTADO_CERRADO and plan.estado != PlanProduccion.ESTADO_CONSUMO_APLICADO:
        plan.estado = PlanProduccion.ESTADO_CONSUMO_APLICADO
        update_fields.append("estado")
    if not plan.consumo_aplicado:
        plan.consumo_aplicado = True
        update_fields.append("consumo_aplicado")
    if plan.consumo_aplicado_en is None:
        plan.consumo_aplicado_en = timezone.now()
        update_fields.append("consumo_aplicado_en")
    if acted_by and getattr(acted_by, "is_authenticated", False) and plan.consumo_aplicado_por_id != acted_by.id:
        plan.consumo_aplicado_por = acted_by
        update_fields.append("consumo_aplicado_por")
    if update_fields:
        plan.save(update_fields=update_fields)
        return True
    return False


def _mark_plan_closed(plan: PlanProduccion, acted_by) -> bool:
    update_fields: list[str] = []
    if plan.estado != PlanProduccion.ESTADO_CERRADO:
        plan.estado = PlanProduccion.ESTADO_CERRADO
        update_fields.append("estado")
    if plan.cerrado_en is None:
        plan.cerrado_en = timezone.now()
        update_fields.append("cerrado_en")
    if acted_by and getattr(acted_by, "is_authenticated", False) and plan.cerrado_por_id != acted_by.id:
        plan.cerrado_por = acted_by
        update_fields.append("cerrado_por")
    if update_fields:
        plan.save(update_fields=update_fields)
        return True
    return False


def _apply_plan_consumption(plan: PlanProduccion, acted_by) -> dict[str, int]:
    explosion = _plan_explosion(plan)
    stats = {
        "insumos_created": 0,
        "insumos_skipped": 0,
        "productos_created": 0,
        "productos_skipped": 0,
    }
    referencia = f"PLAN-PROD:{plan.id}"

    all_existing = True
    for row in explosion["insumos"]:
        cantidad = Decimal(str(row.get("cantidad") or 0))
        if cantidad <= 0:
            continue
        if row.get("is_derived_parent"):
            receta_padre = row.get("parent_recipe")
            if receta_padre is None:
                all_existing = False
                continue
            source_hash = hashlib.sha256(
                f"PLAN_CONSUMO|PRODUCTO|{plan.id}|{receta_padre.id}".encode("utf-8")
            ).hexdigest()
            if MovimientoProductoCedis.objects.filter(source_hash=source_hash).exists():
                stats["productos_skipped"] += 1
            else:
                all_existing = False
        else:
            insumo = row.get("insumo")
            if insumo is None and row.get("insumo_id"):
                insumo = Insumo.objects.filter(pk=row["insumo_id"]).first()
            if insumo is None:
                all_existing = False
                continue
            insumo_canonical = canonical_insumo_by_id(insumo.id) or insumo
            source_hash = hashlib.sha256(
                f"PLAN_CONSUMO|INSUMO|{plan.id}|{insumo_canonical.id}".encode("utf-8")
            ).hexdigest()
            if MovimientoInventario.objects.filter(source_hash=source_hash).exists():
                stats["insumos_skipped"] += 1
            else:
                all_existing = False

    if all_existing and (stats["insumos_skipped"] > 0 or stats["productos_skipped"] > 0):
        with transaction.atomic():
            _mark_plan_consumption_applied(plan, acted_by)
        return stats

    can_apply, reason = _plan_consumo_can_apply(explosion)
    if not can_apply:
        raise ValueError(reason)

    with transaction.atomic():
        for row in explosion["insumos"]:
            cantidad = Decimal(str(row.get("cantidad") or 0))
            if cantidad <= 0:
                continue

            if row.get("is_derived_parent"):
                receta_padre = row.get("parent_recipe")
                if receta_padre is None:
                    continue
                source_hash = hashlib.sha256(
                    f"PLAN_CONSUMO|PRODUCTO|{plan.id}|{receta_padre.id}".encode("utf-8")
                ).hexdigest()
                if MovimientoProductoCedis.objects.filter(source_hash=source_hash).exists():
                    stats["productos_skipped"] += 1
                    continue
                movimiento = MovimientoProductoCedis.objects.create(
                    tipo=MovimientoProductoCedis.TIPO_CONSUMO,
                    receta=receta_padre,
                    cantidad=cantidad,
                    referencia=referencia,
                    source_hash=source_hash,
                )
                _apply_producto_cedis_movimiento(movimiento)
                log_event(
                    acted_by,
                    "CREATE",
                    "recetas.MovimientoProductoCedis",
                    movimiento.id,
                    {
                        "tipo": movimiento.tipo,
                        "receta_id": receta_padre.id,
                        "cantidad": str(cantidad),
                        "referencia": referencia,
                    },
                )
                stats["productos_created"] += 1
                continue

            insumo = row.get("insumo")
            if insumo is None and row.get("insumo_id"):
                insumo = Insumo.objects.filter(pk=row["insumo_id"]).first()
            if insumo is None:
                continue
            insumo_canonical = canonical_insumo_by_id(insumo.id) or insumo
            source_hash = hashlib.sha256(
                f"PLAN_CONSUMO|INSUMO|{plan.id}|{insumo_canonical.id}".encode("utf-8")
            ).hexdigest()
            if MovimientoInventario.objects.filter(source_hash=source_hash).exists():
                stats["insumos_skipped"] += 1
                continue
            movimiento = MovimientoInventario.objects.create(
                tipo=MovimientoInventario.TIPO_CONSUMO,
                insumo=insumo_canonical,
                cantidad=cantidad,
                referencia=referencia,
                source_hash=source_hash,
            )
            existencia, _ = ExistenciaInsumo.objects.get_or_create(insumo=insumo_canonical)
            existencia.stock_actual -= cantidad
            existencia.actualizado_en = timezone.now()
            existencia.save(update_fields=["stock_actual", "actualizado_en"])
            log_event(
                acted_by,
                "CREATE",
                "inventario.MovimientoInventario",
                movimiento.id,
                {
                    "tipo": movimiento.tipo,
                    "insumo_id": insumo_canonical.id,
                    "cantidad": str(cantidad),
                    "referencia": referencia,
                },
            )
            stats["insumos_created"] += 1

        if (
            stats["insumos_created"] > 0
            or stats["productos_created"] > 0
            or stats["insumos_skipped"] > 0
            or stats["productos_skipped"] > 0
        ):
            _mark_plan_consumption_applied(plan, acted_by)

    return stats


def _plan_explosion(plan: PlanProduccion) -> Dict[str, Any]:
    items = (
        plan.items.select_related("receta")
        .prefetch_related(
            "receta__lineas__insumo__unidad_base",
            "receta__lineas__insumo__proveedor_principal",
            "receta__lineas__unidad",
        )
        .order_by("id")
    )

    unit_cost_cache: Dict[int, Decimal | None] = {}
    insumos_map: Dict[int, Dict[str, Any]] = {}
    derived_parent_map: Dict[int, Dict[str, Any]] = {}
    items_detalle: List[Dict[str, Any]] = []
    lineas_sin_cantidad: set[str] = set()
    lineas_sin_costo_unitario: set[str] = set()
    lineas_sin_match = 0

    for item in items:
        multiplicador = Decimal(str(item.cantidad or 0))
        item_total = Decimal("0")
        item_lineas_sin_cantidad = 0
        item_lineas_sin_costo = 0
        item_lineas_sin_match = 0
        item_parent_shortage = False
        item_parent_cost_open = False
        if multiplicador <= 0:
            continue

        for linea in item.receta.lineas.all():
            if not linea.insumo_id:
                lineas_sin_match += 1
                item_lineas_sin_match += 1
                continue

            qty_base = Decimal(str(linea.cantidad or 0))
            if qty_base <= 0:
                lineas_sin_cantidad.add(f"{item.receta.nombre}: {linea.insumo_texto}")
                item_lineas_sin_cantidad += 1
                continue

            qty = qty_base * multiplicador
            if qty <= 0:
                continue

            unit_code = _linea_unit_code(linea)
            unit_cost = _linea_unit_cost(linea, unit_cost_cache)
            costo_linea = Decimal("0")
            if unit_cost is not None and unit_cost > 0:
                costo_linea = qty * unit_cost
            else:
                lineas_sin_costo_unitario.add(f"{item.receta.nombre}: {linea.insumo_texto}")
                item_lineas_sin_costo += 1

            key = linea.insumo_id
            if key not in insumos_map:
                insumo_obj = linea.insumo
                article_class = _insumo_article_class(insumo_obj)
                proveedor_sugerido = "-"
                if insumo_obj.proveedor_principal_id and insumo_obj.proveedor_principal:
                    proveedor_sugerido = insumo_obj.proveedor_principal.nombre
                readiness = _insumo_erp_readiness(insumo_obj)
                master_missing = list(readiness["missing"])
                if insumo_obj.activo and not (insumo_obj.codigo_point or "").strip():
                    master_missing.append("código comercial")
                insumos_map[key] = {
                    "insumo_id": key,
                    "nombre": insumo_obj.nombre,
                    "origen": article_class["label"],
                    "article_class_key": article_class["key"],
                    "article_class_label": article_class["label"],
                    "proveedor_sugerido": proveedor_sugerido,
                    "unidad": unit_code,
                    "cantidad": Decimal("0"),
                    "costo_total": Decimal("0"),
                    "costo_unitario": unit_cost or Decimal("0"),
                    "stock_actual": Decimal("0"),
                    "master_missing": master_missing,
                }

            insumos_map[key]["cantidad"] += qty
            insumos_map[key]["costo_total"] += costo_linea
            item_total += costo_linea

        derived_parent_row = _build_derived_parent_requirement(item.receta, multiplicador)
        if derived_parent_row:
            parent_key = derived_parent_row["parent_recipe_id"]
            if parent_key not in derived_parent_map:
                derived_parent_map[parent_key] = derived_parent_row
            else:
                derived_parent_map[parent_key]["cantidad"] += derived_parent_row["cantidad"]
                derived_parent_map[parent_key]["costo_total"] += derived_parent_row["costo_total"]
                derived_parent_map[parent_key]["costo"] += derived_parent_row["costo"]
                faltante = (
                    Decimal(str(derived_parent_map[parent_key]["cantidad"] or 0))
                    - Decimal(str(derived_parent_map[parent_key]["stock_actual"] or 0))
                )
                derived_parent_map[parent_key]["faltante"] = faltante if faltante > 0 else Decimal("0")
                derived_parent_map[parent_key]["alerta_capacidad"] = (
                    derived_parent_map[parent_key]["faltante"] > 0
                )
            item_total += Decimal(str(derived_parent_row["costo_total"] or 0))
            item_parent_shortage = bool(derived_parent_row["alerta_capacidad"])
            item_parent_cost_open = Decimal(str(derived_parent_row["costo_unitario"] or 0)) <= 0

        workflow_health_label = "Lista para operar"
        workflow_health_tone = "success"
        workflow_action_label = "Ver receta"
        workflow_action_url = reverse("recetas:receta_detail", args=[item.receta.id])
        workflow_next = "Puede entrar a compras y abastecimiento."
        if item_lineas_sin_match > 0:
            workflow_health_label = "Sin artículo estándar"
            workflow_health_tone = "danger"
            workflow_action_label = "Resolver catálogo"
            workflow_action_url = f"{reverse('recetas:matching_pendientes')}?receta={item.receta.id}"
            workflow_next = "Liga componentes al catálogo estándar."
        elif item_lineas_sin_cantidad > 0:
            workflow_health_label = "Sin cantidad"
            workflow_health_tone = "warning"
            workflow_action_label = "Completar receta"
            workflow_next = "Captura cantidades faltantes en la estructura."
        elif item_lineas_sin_costo > 0:
            workflow_health_label = "Sin costo"
            workflow_health_tone = "warning"
            workflow_action_label = "Revisar costos"
            workflow_action_url = f"{reverse('maestros:insumo_list')}?costo_status=sin_costo"
            workflow_next = "Asigna costo vigente al artículo ligado."
        elif item_parent_cost_open:
            workflow_health_label = "Sin costo base padre"
            workflow_health_tone = "warning"
            workflow_action_label = "Abrir padre"
            workflow_action_url = reverse("recetas:receta_detail", args=[derived_parent_row["parent_recipe_id"]])
            workflow_next = "Cierra el costeo del producto padre antes de producir esta presentación."
        elif item_parent_shortage:
            workflow_health_label = "Preparar padre"
            workflow_health_tone = "warning"
            workflow_action_label = "Abrir padre"
            workflow_action_url = reverse("recetas:receta_detail", args=[derived_parent_row["parent_recipe_id"]])
            workflow_next = "Programa disponibilidad del producto padre para cubrir la presentación derivada."

        items_detalle.append(
            {
                "id": item.id,
                "receta": item.receta,
                "cantidad": multiplicador,
                "notas": item.notas,
                "costo_estimado": item_total,
                "lineas_sin_match": item_lineas_sin_match,
                "lineas_sin_cantidad": item_lineas_sin_cantidad,
                "lineas_sin_costo_unitario": item_lineas_sin_costo,
                "workflow_health_label": workflow_health_label,
                "workflow_health_tone": workflow_health_tone,
                "workflow_action_label": workflow_action_label,
                "workflow_action_url": workflow_action_url,
                "workflow_next": workflow_next,
            }
        )

    insumos = sorted(
        [*insumos_map.values(), *derived_parent_map.values()],
        key=lambda x: x["nombre"].lower(),
    )
    existencias_map = {
        e.insumo_id: Decimal(str(e.stock_actual or 0))
        for e in ExistenciaInsumo.objects.filter(insumo_id__in=list(insumos_map.keys()))
    }
    alertas_capacidad = 0
    for row in insumos:
        if row.get("is_derived_parent"):
            if row["alerta_capacidad"]:
                row["workflow_health_label"] = "Preparar padre"
                row["workflow_health_tone"] = "warning"
                row["workflow_action_label"] = "Abrir padre"
                row["workflow_action_url"] = row["detail_url"]
                row["workflow_next"] = "Programa el producto padre para cubrir las presentaciones derivadas del plan."
            if Decimal(str(row["costo_unitario"] or 0)) <= 0:
                row["workflow_health_label"] = "Sin costo base padre"
                row["workflow_health_tone"] = "danger"
                row["workflow_action_label"] = "Abrir padre"
                row["workflow_action_url"] = row["detail_url"]
                row["workflow_next"] = "Cierra el costeo del producto padre antes de liberar producción."
        else:
            row["stock_actual"] = existencias_map.get(row["insumo_id"], Decimal("0"))
            faltante = Decimal(str(row["cantidad"] or 0)) - Decimal(str(row["stock_actual"] or 0))
            row["faltante"] = faltante if faltante > 0 else Decimal("0")
            row["alerta_capacidad"] = row["faltante"] > 0
            row["detail_url"] = reverse("maestros:insumo_update", args=[row["insumo_id"]])
            row["workflow_health_label"] = "Cubierto"
            row["workflow_health_tone"] = "success"
            row["workflow_action_label"] = "Ver artículo"
            row["workflow_action_url"] = row["detail_url"]
            row["workflow_next"] = "Stock suficiente para este plan."
            if Decimal(str(row["costo_unitario"] or 0)) <= 0:
                row["workflow_health_label"] = "Sin costo"
                row["workflow_health_tone"] = "danger"
                row["workflow_action_label"] = "Completar costo"
                row["workflow_next"] = "Asigna costo vigente antes de habilitar compras."
            elif row["origen"] == "Materia prima" and row["proveedor_sugerido"] == "-":
                row["workflow_health_label"] = "Sin proveedor"
                row["workflow_health_tone"] = "warning"
                row["workflow_action_label"] = "Asignar proveedor"
                row["workflow_next"] = "Completa proveedor principal para compras."
            elif row["alerta_capacidad"] and row["origen"] == "Interno":
                row["workflow_health_label"] = "Producir interno"
                row["workflow_health_tone"] = "warning"
                row["workflow_action_label"] = "Revisar receta base"
                row["workflow_action_url"] = f"{reverse('recetas:recetas_list')}?q={urlencode({'q': row['nombre']})[2:]}"
                row["workflow_next"] = "Programa producción interna para cubrir el faltante."
            elif row["alerta_capacidad"]:
                row["workflow_health_label"] = "Comprar"
                row["workflow_health_tone"] = "warning"
                row["workflow_action_label"] = "Ir a compras"
                row["workflow_action_url"] = f"{reverse('compras:solicitudes')}?origen=plan_produccion"
                row["workflow_next"] = "Genera solicitud u orden para cubrir el faltante."
        if row["alerta_capacidad"]:
            alertas_capacidad += 1
    costo_total = sum((row["costo_total"] for row in insumos), Decimal("0"))

    return {
        "items_detalle": items_detalle,
        "insumos": insumos,
        "costo_total": costo_total,
        "lineas_sin_cantidad": sorted(lineas_sin_cantidad),
        "lineas_sin_costo_unitario": sorted(lineas_sin_costo_unitario),
        "lineas_sin_match": lineas_sin_match,
        "alertas_capacidad": alertas_capacidad,
    }


def _plan_vs_pronostico(plan: PlanProduccion) -> Dict[str, Any]:
    periodo = plan.fecha_produccion.strftime("%Y-%m")
    plan_rows = (
        plan.items.values("receta_id", "receta__nombre")
        .annotate(cantidad_plan=Sum("cantidad"))
        .order_by("receta__nombre")
    )
    plan_map = {
        int(r["receta_id"]): {
            "receta_id": int(r["receta_id"]),
            "receta": r["receta__nombre"],
            "cantidad_plan": Decimal(str(r["cantidad_plan"] or 0)),
            "cantidad_pronostico": Decimal("0"),
        }
        for r in plan_rows
    }

    pronosticos_unavailable = False
    try:
        pronosticos = list(
            PronosticoVenta.objects.filter(periodo=periodo).select_related("receta")
        )
    except (OperationalError, ProgrammingError):
        pronosticos = []
        pronosticos_unavailable = True

    for p in pronosticos:
        row = plan_map.get(p.receta_id)
        if row:
            row["cantidad_pronostico"] = Decimal(str(p.cantidad or 0))
        else:
            plan_map[p.receta_id] = {
                "receta_id": p.receta_id,
                "receta": p.receta.nombre,
                "cantidad_plan": Decimal("0"),
                "cantidad_pronostico": Decimal(str(p.cantidad or 0)),
            }

    rows = sorted(plan_map.values(), key=lambda x: x["receta"].lower())
    con_desviacion = 0
    for row in rows:
        row["delta"] = row["cantidad_plan"] - row["cantidad_pronostico"]
        row["sin_pronostico"] = row["cantidad_plan"] > 0 and row["cantidad_pronostico"] <= 0
        row["workflow_health_label"] = "Alineado"
        row["workflow_health_tone"] = "success"
        row["workflow_action_label"] = "Sin acción"
        row["workflow_action_url"] = "#plan-productos"
        row["workflow_next"] = "Plan y pronóstico están alineados."
        if row["sin_pronostico"]:
            row["workflow_health_label"] = "Sin pronóstico"
            row["workflow_health_tone"] = "danger"
            row["workflow_action_label"] = "Cargar pronóstico"
            row["workflow_action_url"] = "#plan-pronosticos"
            row["workflow_next"] = "Importa o captura base de pronóstico para esta receta."
        elif row["delta"] > 0:
            row["workflow_health_label"] = "Sobre plan"
            row["workflow_health_tone"] = "warning"
            row["workflow_action_label"] = "Revisar plan"
            row["workflow_action_url"] = "#plan-productos"
            row["workflow_next"] = "Reduce plan o valida si Ventas debe subir su forecast."
        elif row["delta"] < 0:
            row["workflow_health_label"] = "Bajo plan"
            row["workflow_health_tone"] = "warning"
            row["workflow_action_label"] = "Ajustar forecast"
            row["workflow_action_url"] = "#plan-pronosticos"
            row["workflow_next"] = "Evalúa subir plan o corregir forecast base."
        if row["delta"] != 0:
            con_desviacion += 1

    return {
        "periodo": periodo,
        "rows": rows,
        "total_plan": sum((r["cantidad_plan"] for r in rows), Decimal("0")),
        "total_pronostico": sum((r["cantidad_pronostico"] for r in rows), Decimal("0")),
        "desviaciones": con_desviacion,
        "pronosticos_unavailable": pronosticos_unavailable,
    }


def _periodo_mrp_resumen(
    periodo_mes: str,
    periodo_tipo: str = "mes",
    focus_kind: str = "",
    focus_key: str = "",
) -> Dict[str, Any]:
    periodo = _normalize_periodo_mes(periodo_mes)
    focus_kind_norm = (focus_kind or "").strip().lower()
    focus_key_norm = (focus_key or "").strip().lower()
    if focus_kind_norm not in {"quality", "master", "master_missing", "chain"}:
        focus_kind_norm = ""
        focus_key_norm = ""
    try:
        year, month = periodo.split("-")
        year_i = int(year)
        month_i = int(month)
    except Exception:
        today = timezone.localdate()
        year_i = today.year
        month_i = today.month
        periodo = f"{year_i:04d}-{month_i:02d}"

    plans_qs = PlanProduccion.objects.filter(
        fecha_produccion__year=year_i,
        fecha_produccion__month=month_i,
    ).order_by("fecha_produccion", "id")

    periodo_tipo_norm = (periodo_tipo or "mes").strip().lower()
    if periodo_tipo_norm not in {"mes", "q1", "q2"}:
        periodo_tipo_norm = "mes"
    if periodo_tipo_norm == "q1":
        plans_qs = plans_qs.filter(fecha_produccion__day__lte=15)
    elif periodo_tipo_norm == "q2":
        plans_qs = plans_qs.filter(fecha_produccion__day__gte=16)

    plans = list(plans_qs.only("id", "nombre", "fecha_produccion"))
    if not plans:
        return {
            "periodo": periodo,
            "periodo_tipo": periodo_tipo_norm,
            "planes_count": 0,
            "planes": [],
            "insumos_count": 0,
            "costo_total": Decimal("0"),
            "alertas_capacidad": 0,
            "lineas_sin_match": 0,
            "lineas_sin_cantidad": 0,
            "lineas_sin_costo_unitario": 0,
            "insumos": [],
            "selected_focus_kind": "",
            "selected_focus_key": "",
            "focus_summary": None,
        }

    plan_items_map = {
        row["plan_id"]: int(row["items_count"] or 0)
        for row in (
            PlanProduccionItem.objects.filter(plan_id__in=[p.id for p in plans])
            .values("plan_id")
            .annotate(items_count=Count("id"))
        )
    }

    items = (
        PlanProduccionItem.objects.filter(plan_id__in=[p.id for p in plans])
        .select_related("plan", "receta")
        .prefetch_related(
            "receta__lineas__insumo__unidad_base",
            "receta__lineas__insumo__proveedor_principal",
            "receta__lineas__unidad",
        )
        .order_by("plan__fecha_produccion", "plan_id", "id")
    )

    unit_cost_cache: Dict[int, Decimal | None] = {}
    insumos_map: Dict[int, Dict[str, Any]] = {}
    lineas_sin_cantidad = 0
    lineas_sin_costo_unitario = 0
    lineas_sin_match = 0

    for item in items:
        multiplicador = Decimal(str(item.cantidad or 0))
        if multiplicador <= 0:
            continue
        for linea in item.receta.lineas.all():
            if not linea.insumo_id:
                lineas_sin_match += 1
                continue

            qty_base = Decimal(str(linea.cantidad or 0))
            if qty_base <= 0:
                lineas_sin_cantidad += 1
                continue

            qty = qty_base * multiplicador
            if qty <= 0:
                continue

            unit_code = _linea_unit_code(linea)
            unit_cost = _linea_unit_cost(linea, unit_cost_cache)
            if unit_cost is None or unit_cost <= 0:
                lineas_sin_costo_unitario += 1
                unit_cost = Decimal("0")
            costo_linea = qty * unit_cost

            key = linea.insumo_id
            if key not in insumos_map:
                insumo_obj = linea.insumo
                article_class = _insumo_article_class(insumo_obj)
                proveedor_sugerido = "-"
                if insumo_obj.proveedor_principal_id and insumo_obj.proveedor_principal:
                    proveedor_sugerido = insumo_obj.proveedor_principal.nombre
                readiness = _insumo_erp_readiness(insumo_obj)
                master_missing = list(readiness["missing"])
                if insumo_obj.activo and not (insumo_obj.codigo_point or "").strip():
                    master_missing.append("código comercial")
                insumos_map[key] = {
                    "insumo_id": key,
                    "nombre": insumo_obj.nombre,
                    "origen": article_class["label"],
                    "article_class_key": article_class["key"],
                    "article_class_label": article_class["label"],
                    "proveedor_sugerido": proveedor_sugerido,
                    "unidad": unit_code,
                    "cantidad": Decimal("0"),
                    "costo_total": Decimal("0"),
                    "costo_unitario": unit_cost,
                    "stock_actual": Decimal("0"),
                    "master_missing": master_missing,
                }

            insumos_map[key]["cantidad"] += qty
            insumos_map[key]["costo_total"] += costo_linea
            if insumos_map[key]["costo_unitario"] <= 0 and unit_cost > 0:
                insumos_map[key]["costo_unitario"] = unit_cost

    insumos = sorted(insumos_map.values(), key=lambda x: x["nombre"].lower())
    existencias_map = {
        e.insumo_id: Decimal(str(e.stock_actual or 0))
        for e in ExistenciaInsumo.objects.filter(insumo_id__in=list(insumos_map.keys()))
    }

    alertas_capacidad = 0
    master_incompletos = 0
    for row in insumos:
        row["stock_actual"] = existencias_map.get(row["insumo_id"], Decimal("0"))
        faltante = Decimal(str(row["cantidad"] or 0)) - Decimal(str(row["stock_actual"] or 0))
        row["faltante"] = faltante if faltante > 0 else Decimal("0")
        row["alerta_capacidad"] = row["faltante"] > 0
        if row["alerta_capacidad"]:
            alertas_capacidad += 1
        row["costo_missing"] = Decimal(str(row.get("costo_unitario") or 0)) <= 0
        row["master_incomplete"] = bool(row.get("master_missing"))
        if row["master_incomplete"]:
            master_incompletos += 1

    quality_cards: list[dict[str, Any]] = []
    blocker_detail_rows: list[dict[str, Any]] = []
    master_blocker_groups: dict[str, dict[str, Any]] = {}
    master_blocker_missing_groups: dict[str, dict[str, Any]] = {}
    master_blocker_detail_rows: list[dict[str, Any]] = []

    if lineas_sin_match:
        quality_cards.append(
            {
                "key": "sin_match",
                "label": "Sin artículo estándar",
                "count": lineas_sin_match,
                "tone": "danger",
                "action_label": "Resolver catálogo",
                "action_url": reverse("recetas:matching_pendientes"),
            }
        )
        blocker_detail_rows.append(
            {
                "key": "sin_match",
                "scope": "Datos",
                "name": "Componentes sin artículo estándar",
                "label": "Sin artículo estándar",
                "detail": f"{lineas_sin_match} componente(s) BOM del período siguen sin ligar a un artículo estándar.",
                "tone": "danger",
                "action_label": "Abrir centro de artículos",
                "action_url": reverse("recetas:matching_pendientes"),
            }
        )
    if lineas_sin_cantidad:
        quality_cards.append(
            {
                "key": "sin_cantidad",
                "label": "Sin cantidad",
                "count": lineas_sin_cantidad,
                "tone": "warning",
                "action_label": "Revisar recetas",
                "action_url": reverse("recetas:recetas_list") + "?enterprise_status=incompletas&governance=sin_componentes",
            }
        )
        blocker_detail_rows.append(
            {
                "key": "sin_cantidad",
                "scope": "Datos",
                "name": "Componentes sin cantidad",
                "label": "Sin cantidad",
                "detail": f"{lineas_sin_cantidad} componente(s) del período no tienen cantidad útil para explotar MRP.",
                "tone": "warning",
                "action_label": "Abrir recetas",
                "action_url": reverse("recetas:recetas_list") + "?enterprise_status=incompletas",
            }
        )
    if lineas_sin_costo_unitario:
        quality_cards.append(
            {
                "key": "sin_costo",
                "label": "Sin costo",
                "count": lineas_sin_costo_unitario,
                "tone": "warning",
                "action_label": "Completar costos",
                "action_url": reverse("maestros:insumo_list") + "?costo_status=sin_costo",
            }
        )
        blocker_detail_rows.append(
            {
                "key": "sin_costo",
                "scope": "Datos",
                "name": "Componentes sin costo unitario",
                "label": "Sin costo",
                "detail": f"{lineas_sin_costo_unitario} línea(s) del período no tienen costo unitario vigente.",
                "tone": "warning",
                "action_label": "Abrir maestro",
                "action_url": reverse("maestros:insumo_list") + "?costo_status=sin_costo",
            }
        )
    if alertas_capacidad:
        quality_cards.append(
            {
                "key": "stock_insuficiente",
                "label": "Stock insuficiente",
                "count": alertas_capacidad,
                "tone": "danger",
                "action_label": "Revisar existencias",
                "action_url": reverse("inventario:existencias"),
            }
        )
    if master_incompletos:
        quality_cards.append(
            {
                "key": "maestro_incompleto",
                "label": "Maestro incompleto",
                "count": master_incompletos,
                "tone": "warning",
                "action_label": "Abrir maestro",
                "action_url": reverse("maestros:insumo_list") + "?enterprise_status=incompletos&usage_scope=recipes",
            }
        )

    article_class_groups: dict[str, dict[str, Any]] = {}

    for row in insumos:
        class_key = str(row.get("article_class_key") or Insumo.TIPO_MATERIA_PRIMA)
        class_label = str(row.get("article_class_label") or row.get("origen") or "Artículo")
        group = article_class_groups.setdefault(
            class_key,
            {
                "key": class_key.lower(),
                "class_key": class_key,
                "class_label": class_label,
                "count": 0,
                "faltante_count": 0,
                "cantidad_total": Decimal("0"),
                "costo_total": Decimal("0"),
            },
        )
        group["count"] += 1
        group["cantidad_total"] += Decimal(str(row.get("cantidad") or 0))
        group["costo_total"] += Decimal(str(row.get("costo_total") or 0))
        if row["alerta_capacidad"]:
            blocker_detail_rows.append(
                {
                    "key": "stock_insuficiente",
                    "scope": "Insumo",
                    "name": row["nombre"],
                    "label": "Stock insuficiente",
                    "detail": (
                        f"Requiere {row['cantidad']:.2f} {row['unidad']} y solo hay {row['stock_actual']:.2f}. "
                        f"Faltante estimado {row['faltante']:.2f}."
                    ),
                    "tone": "danger",
                    "action_label": "Ver existencias",
                    "action_url": reverse("inventario:existencias") + f"?q={row['nombre']}",
                }
            )
            group["faltante_count"] += 1
        if row["master_incomplete"]:
            group = master_blocker_groups.setdefault(
                class_key,
                {
                    "class_key": class_key,
                    "class_label": class_label,
                    "count": 0,
                    "missing_totals": defaultdict(int),
                },
            )
            group["count"] += 1
            for missing_label in row.get("master_missing") or []:
                group["missing_totals"][missing_label] += 1

            primary_missing = (row.get("master_missing") or [None])[0]
            action_meta = _enterprise_blocker_action_meta_for_recipes(
                row["nombre"],
                class_key,
                primary_missing,
                insumo_id=getattr(row.get("insumo"), "id", None) or row.get("insumo_id"),
                usage_scope="recipes",
            )
            master_blocker_detail_rows.append(
                {
                    "key": "maestro_incompleto",
                    "class_key": class_key,
                    "class_label": class_label,
                    "name": row["nombre"],
                    "missing": ", ".join(row.get("master_missing") or []),
                    "detail": "Completa el maestro para liberar costeo, MRP y compras del período.",
                    "action_label": action_meta["label"],
                    "action_detail": action_meta["detail"],
                    "action_url": action_meta["url"],
                    "edit_url": action_meta["edit_url"],
                    "missing_field": _missing_field_to_filter_key(primary_missing) or "maestro",
                }
            )
            for missing_label in row.get("master_missing") or []:
                missing_key = _missing_field_to_filter_key(missing_label) or "maestro"
                missing_group = master_blocker_missing_groups.setdefault(
                    missing_key,
                    {
                        "key": missing_key,
                        "missing_label": missing_label,
                        "count": 0,
                        "class_totals": defaultdict(int),
                    },
                )
                missing_group["count"] += 1
                missing_group["class_totals"][class_label] += 1

    master_blocker_class_cards: list[dict[str, Any]] = []
    for group in sorted(master_blocker_groups.values(), key=lambda item: (-item["count"], item["class_label"])):
        dominant_label = ""
        dominant_count = 0
        for missing_label, count in dict(group["missing_totals"]).items():
            if count > dominant_count:
                dominant_label = missing_label
                dominant_count = count
        action_label, action_detail = _enterprise_blocker_label_detail_for_missing_recipes(dominant_label)
        query = {
            "tipo_item": group["class_key"],
            "enterprise_status": "incompletos",
            "usage_scope": "recipes",
        }
        filter_key = _missing_field_to_filter_key(dominant_label)
        if filter_key:
            query["missing_field"] = filter_key
        master_blocker_class_cards.append(
            {
                "key": group["class_key"].lower(),
                "class_label": group["class_label"],
                "count": group["count"],
                "dominant_label": dominant_label or "maestro incompleto",
                "dominant_count": dominant_count,
                "action_label": action_label,
                "action_detail": action_detail,
                "action_url": reverse("maestros:insumo_list") + f"?{urlencode(query)}",
            }
        )
    master_blocker_missing_cards: list[dict[str, Any]] = []
    for missing_group in sorted(master_blocker_missing_groups.values(), key=lambda item: (-item["count"], item["missing_label"])):
        dominant_class_label = ""
        dominant_class_count = 0
        for class_label, count in dict(missing_group["class_totals"]).items():
            if count > dominant_class_count:
                dominant_class_label = class_label
                dominant_class_count = count
        action_label, action_detail = _enterprise_blocker_label_detail_for_missing_recipes(
            missing_group["missing_label"]
        )
        query = {
            "enterprise_status": "incompletos",
            "usage_scope": "recipes",
            "missing_field": missing_group["key"],
        }
        master_blocker_missing_cards.append(
            {
                "key": missing_group["key"],
                "missing_label": missing_group["missing_label"],
                "count": missing_group["count"],
                "dominant_class_label": dominant_class_label or "Artículo",
                "dominant_class_count": dominant_class_count,
                "action_label": action_label,
                "action_detail": action_detail,
                "action_url": reverse("maestros:insumo_list") + f"?{urlencode(query)}",
            }
        )

    if lineas_sin_match or lineas_sin_cantidad or lineas_sin_costo_unitario:
        health_label = "Con bloqueos operativos"
        health_tone = "danger"
        health_detail = "El MRP del período tiene bloqueos de datos que afectan compras, costeo o abastecimiento."
    elif alertas_capacidad or master_incompletos:
        health_label = "Atención operativa"
        health_tone = "warning"
        health_detail = "El MRP del período ya está calculado, pero hay faltantes de stock o artículos incompletos en el maestro."
    else:
        health_label = "Lista para operar"
        health_tone = "success"
        health_detail = "El período no tiene bloqueos críticos de datos ni faltantes operativos visibles."

    article_class_cards = sorted(
        article_class_groups.values(),
        key=lambda item: (-int(item["count"]), item["class_label"]),
    )

    all_insumos = list(insumos)
    all_blocker_rows = list(blocker_detail_rows)
    all_master_rows = list(master_blocker_detail_rows)
    selected_focus_kind = focus_kind_norm
    selected_focus_key = focus_key_norm
    focus_summary = None

    if selected_focus_kind == "quality" and selected_focus_key:
        if selected_focus_key == "maestro_incompleto":
            insumos = [row for row in all_insumos if row.get("master_incomplete")]
            blocker_detail_rows = [row for row in all_blocker_rows if row.get("key") == "maestro_incompleto"]
            master_blocker_detail_rows = list(all_master_rows)
            focus_summary = {
                "label": "Vista enfocada: maestro incompleto",
                "detail": "Mostrando solo artículos del período con faltantes del maestro.",
            }
        else:
            blocker_detail_rows = [row for row in all_blocker_rows if row.get("key") == selected_focus_key]
            master_blocker_detail_rows = []
            if selected_focus_key == "stock_insuficiente":
                insumos = [row for row in all_insumos if row.get("alerta_capacidad")]
                focus_summary = {
                    "label": "Vista enfocada: stock insuficiente",
                    "detail": "Mostrando solo insumos del período con faltante operativo.",
                }
            elif selected_focus_key == "sin_costo":
                insumos = [row for row in all_insumos if row.get("costo_missing")]
                focus_summary = {
                    "label": "Vista enfocada: sin costo",
                    "detail": "Mostrando solo insumos del período sin costo unitario vigente.",
                }
            elif selected_focus_key == "sin_match":
                insumos = list(all_insumos)
                focus_summary = {
                "label": "Vista enfocada: pendientes de integración",
                "detail": "Mostrando el subconjunto de bloqueos de datos por partidas sin ligar al catálogo estándar.",
                }
            elif selected_focus_key == "sin_cantidad":
                insumos = list(all_insumos)
                focus_summary = {
                    "label": "Vista enfocada: sin cantidad",
                    "detail": "Mostrando el subconjunto de bloqueos de datos por líneas sin cantidad útil.",
                }
    elif selected_focus_kind == "master" and selected_focus_key:
        insumos = [
            row for row in all_insumos
            if row.get("master_incomplete") and str(row.get("article_class_key") or "").lower() == selected_focus_key
        ]
        blocker_detail_rows = []
        master_blocker_detail_rows = [
            row for row in all_master_rows if str(row.get("class_key") or "").lower() == selected_focus_key
        ]
        class_card = next(
            (card for card in master_blocker_class_cards if str(card.get("key") or "").lower() == selected_focus_key),
            None,
        )
        focus_summary = {
            "label": f"Vista enfocada: {class_card['class_label'] if class_card else selected_focus_key}",
            "detail": "Mostrando solo artículos del maestro bloqueando MRP para esta clase operativa.",
        }
    elif selected_focus_kind == "master_missing" and selected_focus_key:
        insumos = [
            row
            for row in all_insumos
            if row.get("master_incomplete")
            and selected_focus_key in {
                _missing_field_to_filter_key(missing_label) or "maestro"
                for missing_label in (row.get("master_missing") or [])
            }
        ]
        blocker_detail_rows = []
        master_blocker_detail_rows = [
            row for row in all_master_rows if str(row.get("missing_field") or "").strip().lower() == selected_focus_key
        ]
        missing_card = next(
            (card for card in master_blocker_missing_cards if str(card.get("key") or "").strip().lower() == selected_focus_key),
            None,
        )
        focus_summary = {
            "label": (
                f"Vista enfocada: {missing_card['missing_label']}"
                if missing_card
                else selected_focus_key
            ),
            "detail": "Mostrando solo artículos del maestro bloqueando MRP por este dato faltante.",
        }
    if selected_focus_kind == "master" and selected_focus_key:
        article_class_cards = [card for card in article_class_cards if str(card.get("key") or "") == selected_focus_key]

    return {
        "periodo": periodo,
        "periodo_tipo": periodo_tipo_norm,
        "planes_count": len(plans),
        "planes": [
            {
                "id": p.id,
                "nombre": p.nombre,
                "fecha_produccion": p.fecha_produccion,
                "estado": p.estado,
                "estado_label": p.get_estado_display(),
                "consumo_aplicado": bool(p.consumo_aplicado),
                "items_count": plan_items_map.get(p.id, 0),
            }
            for p in plans
        ],
        "insumos_count": len(insumos),
        "costo_total": sum((row["costo_total"] for row in insumos), Decimal("0")),
        "alertas_capacidad": alertas_capacidad,
        "master_incompletos": master_incompletos,
        "lineas_sin_match": lineas_sin_match,
        "lineas_sin_cantidad": lineas_sin_cantidad,
        "lineas_sin_costo_unitario": lineas_sin_costo_unitario,
        "health_label": health_label,
        "health_tone": health_tone,
        "health_detail": health_detail,
        "quality_cards": quality_cards,
        "article_class_cards": article_class_cards[:6],
        "blocker_detail_rows": blocker_detail_rows[:12],
        "master_blocker_class_cards": master_blocker_class_cards[:6],
        "master_blocker_missing_cards": master_blocker_missing_cards[:6],
        "master_blocker_detail_rows": master_blocker_detail_rows[:12],
        "insumos": insumos,
        "selected_focus_kind": selected_focus_kind,
        "selected_focus_key": selected_focus_key,
        "focus_summary": focus_summary,
    }


def _plan_enterprise_board(
    plan: PlanProduccion | None,
    explosion: Dict[str, Any] | None,
    plan_vs_pronostico: Dict[str, Any] | None,
    mrp_periodo_resumen: Dict[str, Any] | None,
) -> Dict[str, Any] | None:
    if not plan or not explosion:
        return None

    sin_match = int(explosion.get("lineas_sin_match") or 0)
    sin_cantidad = len(explosion.get("lineas_sin_cantidad") or [])
    sin_costo = len(explosion.get("lineas_sin_costo_unitario") or [])
    faltantes = int(explosion.get("alertas_capacidad") or 0)
    desviaciones = 0
    if plan_vs_pronostico and not plan_vs_pronostico.get("pronosticos_unavailable"):
        desviaciones = int(plan_vs_pronostico.get("desviaciones") or 0)

    data_blockers_total = sin_match + sin_cantidad + sin_costo
    blocked_total = data_blockers_total + faltantes + desviaciones
    insumos_por_abastecer = sum(1 for row in explosion.get("insumos") or [] if Decimal(str(row.get("faltante") or 0)) > 0)
    plan_status_label = plan.get_estado_display()
    if plan.estado == PlanProduccion.ESTADO_CERRADO:
        plan_status_tone = "success"
        plan_status_detail = "El plan ya quedó cerrado documental y operativamente."
    elif plan.estado == PlanProduccion.ESTADO_CONSUMO_APLICADO:
        plan_status_tone = "primary"
        plan_status_detail = "El plan ya consumió inventario y está pendiente de cierre formal."
    else:
        plan_status_tone = "warning"
        plan_status_detail = "El plan sigue en preparación operativa."

    blocker_cards: List[Dict[str, Any]] = []
    next_step_cards: List[Dict[str, Any]] = []

    def add_blocker(label: str, count: int, action_label: str, action_url: str, tone: str = "danger") -> None:
        if count <= 0:
            return
        blocker_cards.append(
            {
                "label": label,
                "count": count,
                "tone": tone,
                "action_label": action_label,
                "action_url": action_url,
            }
        )

    def add_step(label: str, count: int, action_label: str, action_url: str, tone: str = "warning") -> None:
        if count <= 0:
            return
        next_step_cards.append(
            {
                "label": label,
                "count": count,
                "tone": tone,
                "action_label": action_label,
                "action_url": action_url,
            }
        )

    matching_url = reverse("recetas:matching_pendientes")
    costeo_url = reverse("maestros:insumo_list")
    inventario_url = reverse("inventario:existencias")
    compras_anchor = "#plan-compras"
    pronostico_anchor = "#plan-vs-pronostico"
    productos_anchor = "#plan-productos"
    insumos_anchor = "#plan-insumos"

    add_blocker("Sin artículo estándar", sin_match, "Resolver catálogo", matching_url)
    add_blocker("Sin cantidad", sin_cantidad, "Revisar recetas", productos_anchor)
    add_blocker("Sin costo", sin_costo, "Completar costos", f"{costeo_url}?costo_status=sin_costo")
    add_blocker("Stock insuficiente", faltantes, "Revisar inventario", inventario_url)
    add_blocker("Desviaciones forecast", desviaciones, "Alinear plan", pronostico_anchor, tone="warning")

    add_step("Resolver catálogo", sin_match, "Abrir centro de artículos", matching_url)
    add_step("Completar cantidades", sin_cantidad, "Revisar componentes BOM", productos_anchor)
    add_step("Completar costeo", sin_costo, "Abrir maestro", f"{costeo_url}?costo_status=sin_costo")
    add_step("Abastecer faltantes", faltantes, "Revisar faltantes", insumos_anchor)
    add_step("Ajustar forecast", desviaciones, "Ver comparativo", pronostico_anchor)

    blocker_detail_rows: List[Dict[str, Any]] = []
    seen_blockers: set[tuple[str, str, str]] = set()
    master_blocker_groups: dict[str, dict[str, Any]] = {}
    master_blocker_detail_rows: List[Dict[str, Any]] = []

    for item in explosion.get("items_detalle") or []:
        label = str(item.get("workflow_health_label") or "")
        if not label or label == "Lista para operar":
            continue
        receta_obj = item.get("receta")
        receta_nombre = getattr(receta_obj, "nombre", "") or "Receta"
        detail = str(item.get("workflow_next") or "")
        dedupe_key = ("receta", receta_nombre, label)
        if dedupe_key in seen_blockers:
            continue
        seen_blockers.add(dedupe_key)
        blocker_detail_rows.append(
            {
                "scope": "Receta",
                "owner": "Producción / Costeo",
                "name": receta_nombre,
                "label": label,
                "detail": detail,
                "tone": item.get("workflow_health_tone") or "warning",
                "action_label": item.get("workflow_action_label") or "Abrir receta",
                "action_url": item.get("workflow_action_url") or productos_anchor,
            }
        )

    for row in explosion.get("insumos") or []:
        label = str(row.get("workflow_health_label") or "")
        if not label or label == "Cubierto":
            continue
        item_name = str(row.get("nombre") or "Insumo")
        detail = str(row.get("workflow_next") or "")
        dedupe_key = ("insumo", item_name, label)
        if dedupe_key in seen_blockers:
            continue
        seen_blockers.add(dedupe_key)
        blocker_detail_rows.append(
            {
                "scope": "Insumo",
                "owner": "Maestros / Compras",
                "name": item_name,
                "label": label,
                "detail": detail,
                "tone": row.get("workflow_health_tone") or "warning",
                "action_label": row.get("workflow_action_label") or "Abrir artículo",
                "action_url": row.get("workflow_action_url") or insumos_anchor,
            }
        )

        missing_fields = list(row.get("master_missing") or [])
        if missing_fields:
            class_key = str(row.get("article_class_key") or Insumo.TIPO_MATERIA_PRIMA)
            class_label = str(row.get("article_class_label") or row.get("origen") or "Artículo")
            group = master_blocker_groups.setdefault(
                class_key,
                {
                    "class_key": class_key,
                    "class_label": class_label,
                    "count": 0,
                    "missing_totals": defaultdict(int),
                    "items": [],
                },
            )
            group["count"] += 1
            group["items"].append(row)
            for missing_label in missing_fields:
                group["missing_totals"][missing_label] += 1

            missing_human = ", ".join(missing_fields)
            primary_missing = missing_fields[0] if missing_fields else None
            action_meta = _enterprise_blocker_action_meta_for_recipes(
                item_name,
                class_key,
                primary_missing,
                insumo_id=getattr(row.get("insumo"), "id", None) or row.get("insumo_id"),
                usage_scope="recipes",
            )
            master_blocker_detail_rows.append(
                {
                    "class_label": class_label,
                    "owner": "Maestros / DG",
                    "name": item_name,
                    "missing": missing_human,
                    "detail": (
                        f"Completa {missing_human} para que el artículo quede listo para compras, "
                        "costeo y abastecimiento."
                    ),
                    "action_label": action_meta["label"],
                    "action_detail": action_meta["detail"],
                    "action_url": action_meta["url"],
                    "edit_url": action_meta["edit_url"],
                }
            )

    for row in (plan_vs_pronostico or {}).get("rows") or []:
        label = str(row.get("workflow_health_label") or "")
        if not label or label == "Alineado":
            continue
        recipe_name = str(row.get("receta") or "Receta")
        detail = "Ajusta el plan o el pronóstico para alinear demanda y producción."
        dedupe_key = ("forecast", recipe_name, label)
        if dedupe_key in seen_blockers:
            continue
        seen_blockers.add(dedupe_key)
        blocker_detail_rows.append(
            {
                "scope": "Forecast",
                "owner": "Ventas / Planeación",
                "name": recipe_name,
                "label": label,
                "detail": detail,
                "tone": row.get("workflow_health_tone") or "warning",
                "action_label": "Ver comparativo",
                "action_url": row.get("workflow_action_url") or pronostico_anchor,
            }
        )

    release_label = "Por validar"
    release_tone = "danger"
    release_detail = "Corrige bloqueos antes de habilitar compras."
    ready_for_purchase_total = 0
    if data_blockers_total == 0:
        ready_for_purchase_total = insumos_por_abastecer
        if ready_for_purchase_total > 0:
            release_label = "Listo para compras"
            release_tone = "success"
            release_detail = "Puedes generar solicitudes y órdenes por faltante."
        else:
            release_label = "Listo / stock cubierto"
            release_tone = "primary"
            release_detail = "No hay faltantes de abastecimiento para este plan."

    period_health = None
    if mrp_periodo_resumen and int(mrp_periodo_resumen.get("planes_count") or 0) > 0:
        period_health = {
            "planes_count": int(mrp_periodo_resumen.get("planes_count") or 0),
            "alertas_capacidad": int(mrp_periodo_resumen.get("alertas_capacidad") or 0),
            "lineas_sin_match": int(mrp_periodo_resumen.get("lineas_sin_match") or 0),
            "lineas_sin_cantidad": int(mrp_periodo_resumen.get("lineas_sin_cantidad") or 0),
            "lineas_sin_costo_unitario": int(mrp_periodo_resumen.get("lineas_sin_costo_unitario") or 0),
            "costo_total": mrp_periodo_resumen.get("costo_total") or Decimal("0"),
            "anchor": "#mrp-periodo",
        }

    master_blocker_class_cards: List[Dict[str, Any]] = []
    master_blocker_missing_groups: dict[str, dict[str, Any]] = {}
    for group in sorted(master_blocker_groups.values(), key=lambda item: (-item["count"], item["class_label"])):
        dominant_label = ""
        dominant_count = 0
        for missing_label, count in dict(group["missing_totals"]).items():
            if count > dominant_count:
                dominant_label = missing_label
                dominant_count = count
        action_label, action_detail = _enterprise_blocker_label_detail_for_missing_recipes(dominant_label)
        query = {
            "tipo_item": group["class_key"],
            "enterprise_status": "incompletos",
            "usage_scope": "recipes",
        }
        filter_key = _missing_field_to_filter_key(dominant_label)
        if filter_key:
            query["missing_field"] = filter_key
        master_blocker_class_cards.append(
            {
                "class_label": group["class_label"],
                "count": group["count"],
                "dominant_label": dominant_label or "maestro incompleto",
                "dominant_count": dominant_count,
                "action_label": action_label,
                "action_detail": action_detail,
                "action_url": f"{costeo_url}?{urlencode(query)}",
            }
        )
        for missing_label, count in dict(group["missing_totals"]).items():
            missing_key = _missing_field_to_filter_key(missing_label) or "maestro"
            missing_group = master_blocker_missing_groups.setdefault(
                missing_key,
                {
                    "missing_key": missing_key,
                    "missing_label": missing_label,
                    "count": 0,
                    "class_totals": defaultdict(int),
                },
            )
            missing_group["count"] += count
            missing_group["class_totals"][group["class_label"]] += count

    master_blocker_missing_cards: List[Dict[str, Any]] = []
    for missing_group in sorted(master_blocker_missing_groups.values(), key=lambda item: (-item["count"], item["missing_label"])):
        dominant_class_label = ""
        dominant_class_count = 0
        for class_label, count in dict(missing_group["class_totals"]).items():
            if count > dominant_class_count:
                dominant_class_label = class_label
                dominant_class_count = count
        action_label, action_detail = _enterprise_blocker_label_detail_for_missing_recipes(
            missing_group["missing_label"]
        )
        query = {
            "enterprise_status": "incompletos",
            "usage_scope": "recipes",
            "missing_field": missing_group["missing_key"],
        }
        master_blocker_missing_cards.append(
            {
                "key": missing_group["missing_key"],
                "missing_label": missing_group["missing_label"],
                "count": missing_group["count"],
                "dominant_class_label": dominant_class_label or "Artículo",
                "dominant_class_count": dominant_class_count,
                "action_label": action_label,
                "action_detail": action_detail,
                "action_url": f"{costeo_url}?{urlencode(query)}",
            }
        )

    return {
        "blocked_total": blocked_total,
        "data_blockers_total": data_blockers_total,
        "ready_for_purchase_total": ready_for_purchase_total,
        "plan_status_label": plan_status_label,
        "plan_status_tone": plan_status_tone,
        "plan_status_detail": plan_status_detail,
        "release_label": release_label,
        "release_tone": release_tone,
        "release_detail": release_detail,
        "next_step_cards": next_step_cards,
        "blocker_cards": blocker_cards,
        "blocker_detail_rows": blocker_detail_rows[:12],
        "master_blocker_class_cards": master_blocker_class_cards[:6],
        "master_blocker_missing_cards": master_blocker_missing_cards[:6],
        "master_blocker_detail_rows": master_blocker_detail_rows[:12],
        "period_health": period_health,
        "purchase_action_url": compras_anchor,
    }


def _plan_document_control(
    plan: PlanProduccion | None,
    *,
    stage_key: str = "auto",
    closure_key: str = "auto",
    handoff_key: str = "auto",
    master_focus_key: str = "auto",
    master_missing_key: str = "auto",
) -> Dict[str, Any] | None:
    if not plan:
        return None

    if stage_key not in {"auto", "solicitudes", "ordenes", "recepciones"}:
        stage_key = "auto"
    if closure_key not in {"auto", "solicitudes_liberadas", "ordenes_sin_bloqueo", "recepciones_aplicadas"}:
        closure_key = "auto"
    if handoff_key not in {"auto", "solicitud_orden", "orden_recepcion", "recepcion_cierre"}:
        handoff_key = "auto"

    plan_scope = f"PLAN_PRODUCCION:{plan.id}"
    plan_focus_base = f"{reverse('recetas:plan_produccion')}?{urlencode({'plan_id': plan.id})}"
    solicitudes_url = f"{reverse('compras:solicitudes')}?{urlencode({'source': 'plan', 'plan_id': plan.id})}"
    ordenes_url = f"{reverse('compras:ordenes')}?{urlencode({'source': 'plan', 'plan_id': plan.id})}"
    recepciones_url = f"{reverse('compras:recepciones')}?{urlencode({'source': 'plan', 'plan_id': plan.id})}"
    solicitudes_qs = SolicitudCompra.objects.filter(area=plan_scope)
    ordenes_qs = OrdenCompra.objects.filter(referencia=plan_scope)
    recepciones_qs = RecepcionCompra.objects.filter(orden__referencia=plan_scope)
    solicitudes_con_insumo = list(
        solicitudes_qs.select_related("insumo__unidad_base", "insumo__proveedor_principal").exclude(insumo__isnull=True)
    )

    solicitudes_total = solicitudes_qs.count()
    ordenes_total = ordenes_qs.count()
    recepciones_total = recepciones_qs.count()

    solicitudes_borrador = solicitudes_qs.filter(estatus=SolicitudCompra.STATUS_BORRADOR).count()
    solicitudes_revision = solicitudes_qs.filter(estatus=SolicitudCompra.STATUS_EN_REVISION).count()
    solicitudes_aprobadas = solicitudes_qs.filter(estatus=SolicitudCompra.STATUS_APROBADA).count()

    ordenes_borrador = ordenes_qs.filter(estatus=OrdenCompra.STATUS_BORRADOR).count()
    ordenes_enviadas = ordenes_qs.filter(estatus=OrdenCompra.STATUS_ENVIADA).count()
    ordenes_confirmadas = ordenes_qs.filter(estatus=OrdenCompra.STATUS_CONFIRMADA).count()
    ordenes_parciales = ordenes_qs.filter(estatus=OrdenCompra.STATUS_PARCIAL).count()
    ordenes_cerradas = ordenes_qs.filter(estatus=OrdenCompra.STATUS_CERRADA).count()

    recepciones_pendientes = recepciones_qs.filter(estatus=RecepcionCompra.STATUS_PENDIENTE).count()
    recepciones_diferencias = recepciones_qs.filter(estatus=RecepcionCompra.STATUS_DIFERENCIAS).count()
    recepciones_cerradas = recepciones_qs.filter(estatus=RecepcionCompra.STATUS_CERRADA).count()

    solicitudes_pendientes_total = solicitudes_borrador + solicitudes_revision
    ordenes_pendientes_total = ordenes_borrador + ordenes_enviadas
    recepciones_abiertas_total = recepciones_pendientes + recepciones_diferencias
    blocked_total = (
        solicitudes_pendientes_total
        + ordenes_pendientes_total
        + recepciones_abiertas_total
    )
    plan_status_label = plan.get_estado_display()
    if plan.estado == PlanProduccion.ESTADO_CERRADO:
        plan_status_tone = "success"
        plan_status_detail = (
            f"Cerrado {timezone.localtime(plan.cerrado_en).strftime('%Y-%m-%d %H:%M')}" if plan.cerrado_en else "El plan quedó cerrado operativamente."
        )
        if plan.cerrado_por:
            plan_status_detail += f" · por {plan.cerrado_por}"
    elif plan.estado == PlanProduccion.ESTADO_CONSUMO_APLICADO:
        plan_status_tone = "primary"
        plan_status_detail = "El plan ya consumió inventario real y está pendiente de cierre formal."
    else:
        plan_status_tone = "warning"
        plan_status_detail = "El plan sigue en preparación y todavía no cierra ejecución operativa."
    consumo_aplicado = bool(plan.consumo_aplicado)
    consumo_aplicado_en = plan.consumo_aplicado_en
    consumo_aplicado_por = plan.consumo_aplicado_por
    consumo_card_tone = "success" if consumo_aplicado else "warning"
    if consumo_aplicado:
        consumo_detail_bits = []
        if consumo_aplicado_en:
            consumo_detail_bits.append(
                f"Aplicado {timezone.localtime(consumo_aplicado_en).strftime('%Y-%m-%d %H:%M')}"
            )
        if consumo_aplicado_por:
            consumo_detail_bits.append(f"por {consumo_aplicado_por}")
        consumo_card_detail = " · ".join(consumo_detail_bits) or "El plan ya descontó inventario real."
    else:
        consumo_card_detail = "El plan sigue en simulación; todavía no descuenta inventario."

    health_label = "Sin arrancar"
    health_tone = "warning"
    health_detail = "Todavía no existen documentos ejecutados para este plan."
    if plan.estado == PlanProduccion.ESTADO_CERRADO:
        health_label = "Plan cerrado"
        health_tone = "success"
        health_detail = "El plan quedó cerrado documental y operativamente."
    elif recepciones_total and recepciones_cerradas == recepciones_total and recepciones_abiertas_total == 0:
        health_label = "Cerrado operativo" if consumo_aplicado else "Cerrado documental"
        health_tone = "success" if consumo_aplicado else "warning"
        health_detail = (
            "El ciclo documental está cerrado y el consumo de inventario ya quedó aplicado."
            if consumo_aplicado
            else "El ciclo documental quedó cerrado, pero el consumo de inventario sigue pendiente."
        )
    elif blocked_total > 0:
        health_label = "Con bloqueos"
        health_tone = "danger"
        health_detail = "Hay documentos pendientes de aprobación, confirmación o cierre."
    elif solicitudes_total or ordenes_total or recepciones_total:
        health_label = "En ejecución"
        health_tone = "primary"
        health_detail = "El plan ya tiene documentos activos y sin bloqueos visibles."

    stage_label = "Sin documentos"
    stage_tone = "warning"
    stage_detail = "Todavía no se han generado documentos de compras para este plan."
    next_action_label = "Generar solicitudes"
    next_action_url = "#plan-compras"

    if solicitudes_total == 0 and ordenes_total == 0 and recepciones_total == 0:
        pass
    elif recepciones_total > 0:
        if recepciones_pendientes > 0 or recepciones_diferencias > 0:
            stage_label = "Recepción en proceso"
            stage_tone = "warning"
            stage_detail = "El plan ya tiene recepciones abiertas; falta validar o cerrar la ejecución documental."
            next_action_label = "Abrir recepciones"
            next_action_url = recepciones_url
        elif recepciones_cerradas == recepciones_total:
            stage_label = "Cerrado"
            stage_tone = "success"
            stage_detail = "Las recepciones del plan ya fueron cerradas y aplicadas."
            next_action_label = "Ver recepciones"
            next_action_url = recepciones_url
    elif ordenes_total > 0:
        stage_label = "Compras en tránsito"
        stage_tone = "primary"
        stage_detail = "El plan ya tiene órdenes activas; el siguiente paso es recepción y cierre."
        next_action_label = "Abrir órdenes"
        next_action_url = ordenes_url
    elif solicitudes_total > 0:
        stage_label = "Solicitudes generadas"
        stage_tone = "primary"
        stage_detail = "El plan ya generó solicitudes; falta aprobarlas y convertirlas en órdenes."
        next_action_label = "Abrir solicitudes"
        next_action_url = solicitudes_url

    document_cards = [
        {
            "label": "Solicitudes",
            "count": solicitudes_total,
            "detail": f"Borrador {solicitudes_borrador} · Revisión {solicitudes_revision} · Aprobadas {solicitudes_aprobadas}",
            "tone": "success" if solicitudes_total and solicitudes_borrador == 0 and solicitudes_revision == 0 else ("primary" if solicitudes_total else "warning"),
            "action_label": "Abrir solicitudes",
            "action_url": solicitudes_url,
        },
        {
            "label": "Órdenes",
            "count": ordenes_total,
            "detail": f"Borrador {ordenes_borrador} · Enviadas {ordenes_enviadas} · Confirmadas {ordenes_confirmadas} · Parciales {ordenes_parciales} · Cerradas {ordenes_cerradas}",
            "tone": "success" if ordenes_total and ordenes_cerradas == ordenes_total else ("primary" if ordenes_total else "warning"),
            "action_label": "Abrir órdenes",
            "action_url": ordenes_url,
        },
        {
            "label": "Recepciones",
            "count": recepciones_total,
            "detail": f"Por validar {recepciones_pendientes} · Diferencias {recepciones_diferencias} · Cerradas {recepciones_cerradas}",
            "tone": "success" if recepciones_total and recepciones_cerradas == recepciones_total else ("warning" if recepciones_total else "primary"),
            "action_label": "Abrir recepciones",
            "action_url": recepciones_url,
        },
        {
            "label": "Consumo inventario",
            "count": 1 if consumo_aplicado else 0,
            "detail": consumo_card_detail,
            "tone": consumo_card_tone,
            "action_label": "Abrir plan",
            "action_url": plan_focus_base,
        },
        {
            "label": "Estado del plan",
            "count": 1 if plan.estado == PlanProduccion.ESTADO_CERRADO else 0,
            "detail": plan_status_detail,
            "tone": plan_status_tone,
            "action_label": "Abrir plan",
            "action_url": plan_focus_base,
        },
    ]

    document_health_cards = [
        {
            "label": "Solicitudes pendientes",
            "count": solicitudes_pendientes_total,
            "detail": f"Borrador {solicitudes_borrador} · Revisión {solicitudes_revision}",
            "tone": "warning" if solicitudes_pendientes_total else "success",
            "action_label": "Abrir solicitudes",
            "action_url": solicitudes_url,
            "key": "solicitudes_pendientes",
        },
        {
            "label": "Órdenes por confirmar",
            "count": ordenes_pendientes_total,
            "detail": f"Borrador {ordenes_borrador} · Enviadas {ordenes_enviadas}",
            "tone": "warning" if ordenes_pendientes_total else "success",
            "action_label": "Abrir órdenes",
            "action_url": ordenes_url,
            "key": "ordenes_pendientes",
        },
        {
            "label": "Recepciones abiertas",
            "count": recepciones_abiertas_total,
            "detail": f"Por validar {recepciones_pendientes} · Diferencias {recepciones_diferencias}",
            "tone": "warning" if recepciones_abiertas_total else "success",
            "action_label": "Abrir recepciones",
            "action_url": recepciones_url,
            "key": "recepciones_abiertas",
        },
        {
            "label": "Documentos cerrados",
            "count": solicitudes_aprobadas + ordenes_cerradas + recepciones_cerradas,
            "detail": f"Solicitudes {solicitudes_aprobadas} · Órdenes {ordenes_cerradas} · Recepciones {recepciones_cerradas}",
            "tone": "primary" if (solicitudes_aprobadas + ordenes_cerradas + recepciones_cerradas) else "warning",
            "action_label": "Ver control",
            "action_url": next_action_url,
            "key": "documentos_cerrados",
        },
        {
            "label": "Consumo inventario",
            "count": 1 if consumo_aplicado else 0,
            "detail": consumo_card_detail,
            "tone": consumo_card_tone,
            "action_label": "Abrir plan",
            "action_url": plan_focus_base,
            "key": "consumo_inventario",
        },
        {
            "label": "Estado del plan",
            "count": 1 if plan.estado == PlanProduccion.ESTADO_CERRADO else 0,
            "detail": plan_status_detail,
            "tone": plan_status_tone,
            "action_label": "Abrir plan",
            "action_url": plan_focus_base,
            "key": "estado_plan",
        },
    ]
    document_stage_rows = [
        {
            "key": "solicitudes",
            "label": "Solicitudes",
            "owner": "Compras / Solicitante",
            "open_count": solicitudes_pendientes_total,
            "closed_count": solicitudes_aprobadas,
            "detail": f"Borrador {solicitudes_borrador} · Revisión {solicitudes_revision} · Aprobadas {solicitudes_aprobadas}",
            "action_label": "Abrir solicitudes",
            "action_url": solicitudes_url,
            "focus_url": f"{plan_focus_base}&stage_key=solicitudes",
            "next_step": (
                "Liberar solicitudes del plan"
                if solicitudes_pendientes_total > 0
                else "Solicitud documental al día"
            ),
            "action_detail": (
                "Revisa borradores y solicitudes en revisión antes de emitir órdenes."
                if solicitudes_pendientes_total > 0
                else "Las solicitudes ya quedaron liberadas para el flujo de compras."
            ),
        },
        {
            "key": "ordenes",
            "label": "Órdenes",
            "owner": "Compras",
            "open_count": ordenes_pendientes_total + ordenes_confirmadas + ordenes_parciales,
            "closed_count": ordenes_cerradas,
            "detail": f"Borrador {ordenes_borrador} · Enviadas {ordenes_enviadas} · Confirmadas {ordenes_confirmadas} · Parciales {ordenes_parciales} · Cerradas {ordenes_cerradas}",
            "action_label": "Abrir órdenes",
            "action_url": ordenes_url,
            "focus_url": f"{plan_focus_base}&stage_key=ordenes",
            "next_step": (
                "Cerrar órdenes abiertas"
                if (ordenes_pendientes_total + ordenes_confirmadas + ordenes_parciales) > 0
                else "Órdenes documentales al día"
            ),
            "action_detail": (
                "Convierte solicitudes aprobadas y da seguimiento a órdenes enviadas, confirmadas o parciales."
                if (ordenes_pendientes_total + ordenes_confirmadas + ordenes_parciales) > 0
                else "Las órdenes del plan ya quedaron cerradas documentalmente."
            ),
        },
        {
            "key": "recepciones",
            "label": "Recepciones",
            "owner": "Almacén / Recepción",
            "open_count": recepciones_abiertas_total,
            "closed_count": recepciones_cerradas,
            "detail": f"Por validar {recepciones_pendientes} · Diferencias {recepciones_diferencias} · Cerradas {recepciones_cerradas}",
            "action_label": "Abrir recepciones",
            "action_url": recepciones_url,
            "focus_url": f"{plan_focus_base}&stage_key=recepciones",
            "next_step": (
                "Cerrar recepciones pendientes"
                if recepciones_abiertas_total > 0
                else "Recepciones documentales al día"
            ),
            "action_detail": (
                "Aplica inventario y resuelve diferencias antes de cerrar el plan."
                if recepciones_abiertas_total > 0
                else "Las recepciones ya quedaron cerradas y aplicadas."
            ),
        },
    ]
    for row in document_stage_rows:
        open_count = int(row.get("open_count") or 0)
        closed_count = int(row.get("closed_count") or 0)
        actionable_total = open_count + closed_count
        progress_pct = int(round((closed_count / actionable_total) * 100)) if actionable_total > 0 else 0
        if actionable_total == 0:
            semaphore_label = "Sin documentos"
            semaphore_tone = "muted"
        elif open_count > 0 and closed_count == 0:
            semaphore_label = "Rojo"
            semaphore_tone = "danger"
        elif open_count > 0:
            semaphore_label = "Amarillo"
            semaphore_tone = "warning" if row["label"] != "Órdenes" else "primary"
        else:
            semaphore_label = "Verde"
            semaphore_tone = "success"
        row["total_count"] = actionable_total
        row["progress_pct"] = progress_pct
        row["semaphore_label"] = semaphore_label
        row["semaphore_tone"] = semaphore_tone
    document_blocker_rows: list[dict[str, Any]] = []
    for solicitud in solicitudes_qs.filter(
        estatus__in=[SolicitudCompra.STATUS_BORRADOR, SolicitudCompra.STATUS_EN_REVISION]
    ).order_by("fecha_requerida", "folio")[:8]:
        document_blocker_rows.append(
            {
                "scope": "Solicitud",
                "folio": solicitud.folio,
                "status": solicitud.get_estatus_display(),
                "detail": f"{solicitud.insumo.nombre if solicitud.insumo_id else 'Sin artículo'} · {solicitud.solicitante}",
                "action_label": "Abrir solicitudes",
                "action_url": f"{solicitudes_url}&q={urlencode({'q': solicitud.folio})[2:]}",
            }
        )
    for orden in ordenes_qs.filter(
        estatus__in=[OrdenCompra.STATUS_BORRADOR, OrdenCompra.STATUS_ENVIADA]
    ).order_by("fecha_emision", "folio")[:8]:
        document_blocker_rows.append(
            {
                "scope": "Orden",
                "folio": orden.folio,
                "status": orden.get_estatus_display(),
                "detail": f"{orden.proveedor.nombre if orden.proveedor_id else 'Sin proveedor'} · monto {Decimal(str(orden.monto_estimado or 0)):.2f}",
                "action_label": "Abrir órdenes",
                "action_url": f"{ordenes_url}&q={urlencode({'q': orden.folio})[2:]}",
            }
        )
    for recepcion in recepciones_qs.filter(
        estatus__in=[RecepcionCompra.STATUS_PENDIENTE, RecepcionCompra.STATUS_DIFERENCIAS]
    ).order_by("fecha_recepcion", "folio")[:8]:
        document_blocker_rows.append(
            {
                "scope": "Recepción",
                "folio": recepcion.folio,
                "status": recepcion.get_estatus_display(),
                "detail": f"{recepcion.orden.folio if recepcion.orden_id else 'Sin orden'} · conformidad {Decimal(str(recepcion.conformidad_pct or 0)):.0f}%",
                "action_label": "Abrir recepciones",
                "action_url": f"{recepciones_url}&q={urlencode({'q': recepcion.folio})[2:]}",
            }
        )

    master_blocker_groups: dict[str, dict[str, Any]] = {}
    master_blocker_missing_groups: dict[str, dict[str, Any]] = {}
    master_blocker_detail_rows: List[Dict[str, Any]] = []
    for solicitud in solicitudes_con_insumo:
        insumo = solicitud.insumo
        readiness = _insumo_erp_readiness(insumo)
        missing_fields = list(readiness["missing"])
        if insumo.activo and not (insumo.codigo_point or "").strip():
            missing_fields.append("código comercial")
        if not missing_fields:
            continue

        article_class = _insumo_article_class(insumo)
        class_key = article_class["key"]
        class_label = article_class["label"]
        blocker_group = master_blocker_groups.setdefault(
            class_key,
            {
                "class_key": class_key,
                "class_label": class_label,
                "count": 0,
                "missing_totals": defaultdict(int),
            },
        )
        blocker_group["count"] += 1
        for missing_label in missing_fields:
            blocker_group["missing_totals"][missing_label] += 1

        primary_missing = missing_fields[0]
        action_meta = _enterprise_blocker_action_meta_for_recipes(
            insumo.nombre,
            class_key,
            primary_missing,
            insumo_id=insumo.id,
            usage_scope="recipes",
        )
        master_blocker_detail_rows.append(
            {
                "class_key": class_key,
                "class_label": class_label,
                "owner": "Maestros / DG",
                "name": insumo.nombre,
                "missing": ", ".join(missing_fields),
                "detail": "Completa el maestro para liberar solicitudes, compras y costeo del plan.",
                "action_label": action_meta["label"],
                "action_detail": action_meta["detail"],
                "action_url": action_meta["url"],
                "edit_url": action_meta["edit_url"],
                "missing_field": _missing_field_to_filter_key(primary_missing) or "maestro",
            }
        )
        for missing_label in missing_fields:
            missing_key = _missing_field_to_filter_key(missing_label) or "maestro"
            missing_group = master_blocker_missing_groups.setdefault(
                missing_key,
                {
                    "key": missing_key,
                    "missing_label": missing_label,
                    "count": 0,
                    "class_totals": defaultdict(int),
                },
            )
            missing_group["count"] += 1
            missing_group["class_totals"][class_label] += 1

    master_blocker_class_cards: List[Dict[str, Any]] = []
    for group in sorted(master_blocker_groups.values(), key=lambda item: (-item["count"], item["class_label"])):
        dominant_label = ""
        dominant_count = 0
        for missing_label, count in dict(group["missing_totals"]).items():
            if count > dominant_count:
                dominant_label = missing_label
                dominant_count = count
        action_label, action_detail = _enterprise_blocker_label_detail_for_missing_recipes(dominant_label)
        query = {
            "tipo_item": group["class_key"],
            "enterprise_status": "incompletos",
            "usage_scope": "recipes",
        }
        filter_key = _missing_field_to_filter_key(dominant_label)
        if filter_key:
            query["missing_field"] = filter_key
        master_blocker_class_cards.append(
            {
                "key": group["class_key"].lower(),
                "class_label": group["class_label"],
                "count": group["count"],
                "dominant_label": dominant_label or "maestro incompleto",
                "dominant_count": dominant_count,
                "action_label": action_label,
                "action_detail": action_detail,
                "action_url": reverse("maestros:insumo_list") + f"?{urlencode(query)}",
            }
        )

    master_blocker_missing_cards: List[Dict[str, Any]] = []
    for missing_group in sorted(master_blocker_missing_groups.values(), key=lambda item: (-item["count"], item["missing_label"])):
        dominant_class_label = ""
        dominant_class_count = 0
        for class_label, count in dict(missing_group["class_totals"]).items():
            if count > dominant_class_count:
                dominant_class_label = class_label
                dominant_class_count = count
        action_label, action_detail = _enterprise_blocker_label_detail_for_missing_recipes(
            missing_group["missing_label"]
        )
        query = {
            "enterprise_status": "incompletos",
            "usage_scope": "recipes",
            "missing_field": missing_group["key"],
        }
        master_blocker_missing_cards.append(
            {
                "key": missing_group["key"],
                "missing_label": missing_group["missing_label"],
                "count": missing_group["count"],
                "dominant_class_label": dominant_class_label or "Artículo",
                "dominant_class_count": dominant_class_count,
                "action_label": action_label,
                "action_detail": action_detail,
                "action_url": reverse("maestros:insumo_list") + f"?{urlencode(query)}",
            }
        )
    active_scope = "Solicitud"
    active_scope_key = "solicitudes"
    active_scope_label = "Solicitudes"
    active_scope_url = solicitudes_url
    active_scope_detail = "Genera y libera solicitudes del plan antes de abrir compras."
    active_stage_open = solicitudes_pendientes_total
    active_stage_closed = solicitudes_aprobadas
    if stage_key == "recepciones":
        active_scope = "Recepción"
        active_scope_key = "recepciones"
        active_scope_label = "Recepciones"
        active_scope_url = recepciones_url
        active_scope_detail = (
            "La etapa activa está en recepción; valida diferencias y cierra el documento."
            if recepciones_abiertas_total > 0
            else "Las recepciones del plan ya quedaron cerradas."
        )
        active_stage_open = recepciones_abiertas_total
        active_stage_closed = recepciones_cerradas
    elif stage_key == "ordenes":
        active_scope = "Orden"
        active_scope_key = "ordenes"
        active_scope_label = "Órdenes"
        active_scope_url = ordenes_url
        active_scope_detail = (
            "La etapa activa está en órdenes; confirma o cierra compras antes de recepción."
            if (ordenes_pendientes_total + ordenes_confirmadas + ordenes_parciales) > 0
            else "Las órdenes del plan ya están cerradas."
        )
        active_stage_open = ordenes_pendientes_total + ordenes_confirmadas + ordenes_parciales
        active_stage_closed = ordenes_cerradas
    elif stage_key == "solicitudes":
        active_scope = "Solicitud"
        active_scope_key = "solicitudes"
        active_scope_label = "Solicitudes"
        active_scope_url = solicitudes_url
        active_scope_detail = (
            "La etapa activa está en solicitudes; apruébalas y conviértelas en órdenes."
            if solicitudes_pendientes_total > 0
            else "Las solicitudes del plan ya fueron liberadas."
        )
    elif recepciones_total > 0:
        active_scope = "Recepción"
        active_scope_key = "recepciones"
        active_scope_label = "Recepciones"
        active_scope_url = recepciones_url
        active_scope_detail = (
            "La etapa activa está en recepción; valida diferencias y cierra el documento."
            if recepciones_abiertas_total > 0
            else "Las recepciones del plan ya quedaron cerradas."
        )
        active_stage_open = recepciones_abiertas_total
        active_stage_closed = recepciones_cerradas
    elif ordenes_total > 0:
        active_scope = "Orden"
        active_scope_key = "ordenes"
        active_scope_label = "Órdenes"
        active_scope_url = ordenes_url
        active_scope_detail = (
            "La etapa activa está en órdenes; confirma o cierra compras antes de recepción."
            if (ordenes_pendientes_total + ordenes_confirmadas + ordenes_parciales) > 0
            else "Las órdenes del plan ya están cerradas."
        )
        active_stage_open = ordenes_pendientes_total + ordenes_confirmadas + ordenes_parciales
        active_stage_closed = ordenes_cerradas
    elif solicitudes_total > 0:
        active_scope = "Solicitud"
        active_scope_key = "solicitudes"
        active_scope_label = "Solicitudes"
        active_scope_url = solicitudes_url
        active_scope_detail = (
            "La etapa activa está en solicitudes; apruébalas y conviértelas en órdenes."
            if solicitudes_pendientes_total > 0
            else "Las solicitudes del plan ya fueron liberadas."
        )

    stage_blocker_rows = [row for row in document_blocker_rows if row["scope"] == active_scope][:6]
    stage_focus = {
        "scope": active_scope,
        "key": active_scope_key,
        "label": active_scope_label,
        "detail": active_scope_detail,
        "blocked_count": len(stage_blocker_rows),
        "open_count": active_stage_open,
        "closed_count": active_stage_closed,
        "action_label": f"Abrir {active_scope_label.lower()}",
        "action_url": active_scope_url,
        "blocker_rows": stage_blocker_rows,
        "summary": (
            f"{len(stage_blocker_rows)} bloqueos priorizados en {active_scope_label.lower()}."
            if stage_blocker_rows
            else f"{active_scope_label} sin bloqueos priorizados."
        ),
        "tone": (
            "danger"
            if stage_blocker_rows
            else "warning"
            if active_stage_open > 0
            else "success"
            if active_stage_closed > 0
            else "primary"
        ),
    }
    pipeline_steps = [
        {
            "label": "Solicitudes",
            "owner": "Compras / Solicitante",
            "count": solicitudes_total,
            "open": solicitudes_pendientes_total,
            "blocked": 0,
            "closed": solicitudes_aprobadas,
            "status_label": (
                "Sin generar"
                if solicitudes_total == 0
                else ("Por atender" if solicitudes_pendientes_total > 0 else "Liberadas")
            ),
            "tone": (
                "warning"
                if solicitudes_total == 0 or solicitudes_pendientes_total > 0
                else "success"
            ),
            "detail": (
                "Genera solicitudes del plan."
                if solicitudes_total == 0
                else (
                    f"Borrador {solicitudes_borrador} · Revisión {solicitudes_revision}"
                    if solicitudes_pendientes_total > 0
                    else f"Aprobadas {solicitudes_aprobadas}"
                )
            ),
            "action_label": "Abrir solicitudes",
            "action_url": solicitudes_url,
            "action_detail": (
                "Aprueba o termina la captura para convertir el plan en solicitudes liberadas."
                if solicitudes_total == 0 or solicitudes_pendientes_total > 0
                else "La etapa de solicitudes ya quedó liberada para compras."
            ),
            "next_step": (
                "Liberar solicitudes"
                if solicitudes_total == 0 or solicitudes_pendientes_total > 0
                else "Etapa cerrada"
            ),
        },
        {
            "label": "Órdenes",
            "owner": "Compras",
            "count": ordenes_total,
            "open": ordenes_pendientes_total + ordenes_confirmadas + ordenes_parciales,
            "blocked": 0,
            "closed": ordenes_cerradas,
            "status_label": (
                "Sin emitir"
                if ordenes_total == 0
                else ("Abiertas" if (ordenes_pendientes_total + ordenes_confirmadas + ordenes_parciales) > 0 else "Cerradas")
            ),
            "tone": (
                "warning"
                if ordenes_total == 0
                else ("primary" if (ordenes_pendientes_total + ordenes_confirmadas + ordenes_parciales) > 0 else "success")
            ),
            "detail": (
                "Convierte solicitudes en órdenes."
                if ordenes_total == 0
                else (
                    f"Borrador {ordenes_borrador} · Enviadas {ordenes_enviadas} · Confirmadas {ordenes_confirmadas} · Parciales {ordenes_parciales}"
                    if (ordenes_pendientes_total + ordenes_confirmadas + ordenes_parciales) > 0
                    else f"Cerradas {ordenes_cerradas}"
                )
            ),
            "action_label": "Abrir órdenes",
            "action_url": ordenes_url,
            "action_detail": (
                "Emite y cierra órdenes hasta dejar el abastecimiento documental en verde."
                if ordenes_total == 0 or (ordenes_pendientes_total + ordenes_confirmadas + ordenes_parciales) > 0
                else "La etapa de órdenes ya quedó cerrada."
            ),
            "next_step": (
                "Emitir o cerrar órdenes"
                if ordenes_total == 0 or (ordenes_pendientes_total + ordenes_confirmadas + ordenes_parciales) > 0
                else "Etapa cerrada"
            ),
        },
        {
            "label": "Recepciones",
            "owner": "Almacén / Recepción",
            "count": recepciones_total,
            "open": recepciones_abiertas_total,
            "blocked": blocked_total,
            "closed": recepciones_cerradas,
            "status_label": (
                "Sin recibir"
                if recepciones_total == 0
                else ("Abiertas" if recepciones_abiertas_total > 0 else "Recepciones al día")
            ),
            "tone": (
                "warning"
                if recepciones_total == 0
                else ("warning" if recepciones_abiertas_total > 0 else "success")
            ),
            "detail": (
                "Recibe y cierra compras del plan."
                if recepciones_total == 0
                else (
                    f"Por validar {recepciones_pendientes} · Diferencias {recepciones_diferencias}"
                    if recepciones_abiertas_total > 0
                    else f"Cerradas {recepciones_cerradas}"
                )
            ),
            "action_label": "Abrir recepciones",
            "action_url": recepciones_url,
            "action_detail": (
                "Aplica inventario y resuelve diferencias para cerrar la recepción documental."
                if recepciones_total == 0 or recepciones_abiertas_total > 0
                else "La etapa de recepciones ya quedó al día."
            ),
            "next_step": (
                "Cerrar recepciones"
                if recepciones_total == 0 or recepciones_abiertas_total > 0
                else "Etapa cerrada"
            ),
        },
    ]
    for step in pipeline_steps:
        count = int(step.get("count") or 0)
        blocked = int(step.get("blocked") or 0)
        closed = int(step.get("closed") or 0)
        open_items = int(step.get("open") or 0)
        actionable_total = blocked + open_items + closed
        progress_pct = int(round((closed / actionable_total) * 100)) if actionable_total > 0 else 0
        if blocked > 0:
            semaphore_label = "Rojo"
            semaphore_tone = "danger"
        elif open_items > 0:
            semaphore_label = "Amarillo"
            semaphore_tone = "warning" if step["label"] != "Órdenes" else "primary"
        elif count > 0:
            semaphore_label = "Verde"
            semaphore_tone = "success"
        else:
            semaphore_label = "Sin documentos"
            semaphore_tone = "muted"
        step["progress_pct"] = progress_pct
        step["semaphore_label"] = semaphore_label
        step["semaphore_tone"] = semaphore_tone
        if step["label"] == "Solicitudes":
            if count == 0:
                step["action_label"] = "Generar solicitudes"
                step["action_detail"] = "Todavía no se abrió el flujo documental del plan."
                step["action_url"] = "#plan-compras"
            elif open_items > 0:
                step["action_label"] = "Liberar solicitudes"
                step["action_detail"] = "Aprueba o termina la captura para convertirlas en órdenes."
                step["action_url"] = f"{solicitudes_url}&estatus={SolicitudCompra.STATUS_BORRADOR}"
            else:
                step["action_label"] = "Abrir solicitudes"
                step["action_detail"] = "La etapa de solicitudes ya quedó liberada."
                step["action_url"] = solicitudes_url
        elif step["label"] == "Órdenes":
            if count == 0:
                step["action_label"] = "Emitir órdenes"
                step["action_detail"] = "Convierte solicitudes liberadas en órdenes de compra."
                step["action_url"] = solicitudes_url
            elif open_items > 0:
                step["action_label"] = "Dar seguimiento a órdenes"
                step["action_detail"] = "Completa emisión, envío o confirmación antes de recibir."
                step["action_url"] = ordenes_url
            else:
                step["action_label"] = "Abrir órdenes"
                step["action_detail"] = "La etapa de órdenes ya está cerrada."
                step["action_url"] = ordenes_url
        else:
            if count == 0:
                step["action_label"] = "Preparar recepciones"
                step["action_detail"] = "Las recepciones aún no existen porque falta avanzar compras."
                step["action_url"] = ordenes_url
            elif open_items > 0:
                step["action_label"] = "Cerrar recepciones"
                step["action_detail"] = "Aplica inventario y resuelve diferencias para cerrar el plan."
                step["action_url"] = (
                    f"{recepciones_url}&estatus={RecepcionCompra.STATUS_PENDIENTE}"
                    if recepciones_pendientes > 0
                    else f"{recepciones_url}&estatus={RecepcionCompra.STATUS_DIFERENCIAS}"
                    if recepciones_diferencias > 0
                    else recepciones_url
                )
            else:
                step["action_label"] = "Abrir recepciones"
                step["action_detail"] = "La etapa de recepción ya quedó cerrada y aplicada."
                step["action_url"] = recepciones_url
    purchase_gate = {
        "label": "Bloqueado",
        "tone": "danger",
        "detail": "Corrige bloqueos documentales antes de seguir con compras.",
        "cta_label": next_action_label,
        "cta_url": next_action_url,
    }
    if blocked_total == 0:
        if solicitudes_total == 0:
            purchase_gate = {
                "label": "Listo para generar",
                "tone": "success",
                "detail": "El plan puede generar solicitudes y abrir el flujo de compras.",
                "cta_label": "Generar solicitudes",
                "cta_url": "#plan-compras",
            }
        elif ordenes_total == 0:
            purchase_gate = {
                "label": "Listo para ordenar",
                "tone": "primary",
                "detail": "Las solicitudes ya están liberadas; el siguiente paso es emitir órdenes.",
                "cta_label": "Abrir solicitudes",
                "cta_url": solicitudes_url,
            }
        elif recepciones_total == 0:
            purchase_gate = {
                "label": "Compras en tránsito",
                "tone": "primary",
                "detail": "Las órdenes están activas; falta registrar recepciones.",
                "cta_label": "Abrir órdenes",
                "cta_url": ordenes_url,
            }
        else:
            purchase_gate = {
                "label": "Cierre en curso",
                "tone": "primary" if recepciones_abiertas_total else "success",
                "detail": (
                    "El plan tiene recepciones abiertas; valida diferencias y cierra."
                    if recepciones_abiertas_total
                    else "Las recepciones quedaron cerradas para este plan."
                ),
                "cta_label": "Abrir recepciones",
                "cta_url": recepciones_url,
            }

    closure_checks = [
        {
            "key": "solicitudes_liberadas",
            "label": "Solicitudes liberadas",
            "is_ready": solicitudes_pendientes_total == 0,
            "detail": "Sin solicitudes en borrador o revisión para este plan.",
            "action_label": (
                "Liberar solicitudes pendientes"
                if solicitudes_pendientes_total
                else "Solicitudes lista"
            ),
            "action_url": (
                f"{solicitudes_url}&estatus={SolicitudCompra.STATUS_BORRADOR}"
                if solicitudes_pendientes_total
                else solicitudes_url
            ),
            "action_detail": (
                "Aprueba o termina la captura antes de emitir órdenes."
                if solicitudes_pendientes_total
                else "La etapa de solicitudes ya quedó liberada."
            ),
            "focus_url": f"{plan_focus_base}&closure_key=solicitudes_liberadas",
        },
        {
            "key": "ordenes_sin_bloqueo",
            "label": "Órdenes sin bloqueo",
            "is_ready": ordenes_pendientes_total == 0,
            "detail": "Sin órdenes en borrador o enviadas pendientes de confirmación.",
            "action_label": (
                "Corregir órdenes pendientes"
                if ordenes_pendientes_total
                else "Órdenes lista"
            ),
            "action_url": (
                f"{ordenes_url}&estatus={OrdenCompra.STATUS_BORRADOR}"
                if ordenes_pendientes_total
                else ordenes_url
            ),
            "action_detail": (
                "Completa proveedor, emisión o confirmación antes de pasar a recepción."
                if ordenes_pendientes_total
                else "Las órdenes ya cumplen condiciones para seguir o cerrar."
            ),
            "focus_url": f"{plan_focus_base}&closure_key=ordenes_sin_bloqueo",
        },
        {
            "key": "recepciones_aplicadas",
            "label": "Recepciones aplicadas",
            "is_ready": recepciones_abiertas_total == 0,
            "detail": "Sin recepciones pendientes o con diferencias para este plan.",
            "action_label": (
                "Cerrar recepciones abiertas"
                if recepciones_abiertas_total
                else "Recepciones al día"
            ),
            "action_url": (
                f"{recepciones_url}&estatus={RecepcionCompra.STATUS_PENDIENTE}"
                if recepciones_abiertas_total and recepciones_pendientes > 0
                else f"{recepciones_url}&estatus={RecepcionCompra.STATUS_DIFERENCIAS}"
                if recepciones_abiertas_total
                else recepciones_url
            ),
            "action_detail": (
                "Aplica inventario, valida conformidad y resuelve diferencias."
                if recepciones_abiertas_total
                else "Las recepciones ya quedaron cerradas y aplicadas."
            ),
            "focus_url": f"{plan_focus_base}&closure_key=recepciones_aplicadas",
        },
    ]
    closure_ready_count = sum(1 for item in closure_checks if item["is_ready"])
    closure_total = len(closure_checks)
    closure_progress_pct = int(round((closure_ready_count / closure_total) * 100)) if closure_total else 0
    closure_summary = {
        "label": "Cierre documental listo" if closure_ready_count == closure_total and closure_total else "Cierre documental pendiente",
        "tone": (
            "success"
            if closure_ready_count == closure_total and closure_total
            else "warning"
            if closure_ready_count > 0
            else "danger"
        ),
        "ready_count": closure_ready_count,
        "pending_count": max(closure_total - closure_ready_count, 0),
        "progress_pct": closure_progress_pct,
        "detail": (
            "Solicitudes, órdenes y recepciones ya cumplen el criterio de cierre del plan."
            if closure_ready_count == closure_total and closure_total
            else "Todavía hay criterios de cierre abiertos antes de considerar el plan documentalmente estable."
        ),
    }
    handoff_checks = [
        {
            "key": "solicitud_orden",
            "label": "Solicitud → Orden",
            "is_ready": solicitudes_pendientes_total == 0 and ordenes_total >= solicitudes_aprobadas,
            "ready_count": max(solicitudes_aprobadas - ordenes_total, 0),
            "blocked_count": solicitudes_pendientes_total,
            "detail": "Toda solicitud aprobada del plan debe quedar convertida en orden o justificar su excepción.",
            "action_label": (
                "Liberar solicitudes"
                if solicitudes_pendientes_total
                else "Emitir órdenes"
                if solicitudes_aprobadas > ordenes_total
                else "Handoff completo"
            ),
            "action_url": (
                f"{solicitudes_url}&estatus={SolicitudCompra.STATUS_BORRADOR}"
                if solicitudes_pendientes_total
                else solicitudes_url
                if solicitudes_aprobadas > ordenes_total
                else solicitudes_url
            ),
            "action_detail": (
                "Primero libera solicitudes pendientes antes de emitir órdenes."
                if solicitudes_pendientes_total
                else "Aún hay solicitudes aprobadas sin orden emitida."
                if solicitudes_aprobadas > ordenes_total
                else "No quedan solicitudes aprobadas pendientes de pasar a orden."
            ),
            "focus_url": f"{plan_focus_base}&handoff_key=solicitud_orden",
        },
        {
            "key": "orden_recepcion",
            "label": "Orden → Recepción",
            "is_ready": ordenes_pendientes_total == 0 and recepciones_total >= (ordenes_confirmadas + ordenes_parciales + ordenes_cerradas),
            "ready_count": max((ordenes_confirmadas + ordenes_parciales + ordenes_cerradas) - recepciones_total, 0),
            "blocked_count": ordenes_pendientes_total,
            "detail": "Toda orden confirmada debe abrir recepción para sostener trazabilidad de compras.",
            "action_label": (
                "Corregir órdenes"
                if ordenes_pendientes_total
                else "Abrir recepciones"
                if recepciones_total < (ordenes_confirmadas + ordenes_parciales + ordenes_cerradas)
                else "Handoff completo"
            ),
            "action_url": (
                f"{ordenes_url}&estatus={OrdenCompra.STATUS_BORRADOR}"
                if ordenes_pendientes_total
                else ordenes_url
            ),
            "action_detail": (
                "Completa envío o confirmación antes de pasar a recepción."
                if ordenes_pendientes_total
                else "Todavía hay órdenes confirmadas sin recepción asociada."
                if recepciones_total < (ordenes_confirmadas + ordenes_parciales + ordenes_cerradas)
                else "No quedan órdenes confirmadas pendientes de recepción."
            ),
            "focus_url": f"{plan_focus_base}&handoff_key=orden_recepcion",
        },
        {
            "key": "recepcion_cierre",
            "label": "Recepción → Cierre",
            "is_ready": recepciones_abiertas_total == 0,
            "ready_count": recepciones_abiertas_total,
            "blocked_count": recepciones_abiertas_total,
            "detail": "Toda recepción debe cerrarse y aplicar inventario antes de cerrar documentalmente el plan.",
            "action_label": (
                "Cerrar recepciones"
                if recepciones_abiertas_total
                else "Handoff completo"
            ),
            "action_url": (
                f"{recepciones_url}&estatus={RecepcionCompra.STATUS_PENDIENTE}"
                if recepciones_pendientes > 0
                else f"{recepciones_url}&estatus={RecepcionCompra.STATUS_DIFERENCIAS}"
                if recepciones_diferencias > 0
                else recepciones_url
            ),
            "action_detail": (
                "Aplica inventario y resuelve diferencias antes de concluir el flujo."
                if recepciones_abiertas_total
                else "Las recepciones ya quedaron cerradas y aplicadas."
            ),
            "focus_url": f"{plan_focus_base}&handoff_key=recepcion_cierre",
        },
    ]
    total_handoff_blocked = sum(int(item.get("blocked_count") or 0) for item in handoff_checks)
    handoff_ready_count = 0 if total_handoff_blocked > 0 else sum(1 for item in handoff_checks if item["is_ready"])
    handoff_total = len(handoff_checks)
    handoff_progress_pct = int(round((handoff_ready_count / handoff_total) * 100)) if handoff_total else 0
    handoff_summary = {
        "label": (
            "Entregas entre etapas listas"
            if handoff_ready_count == handoff_total and handoff_total
            else "Entregas entre etapas pendientes"
        ),
        "tone": (
            "success"
            if handoff_ready_count == handoff_total and handoff_total
            else "danger"
            if total_handoff_blocked > 0
            else "warning"
        ),
        "ready_count": handoff_ready_count,
        "pending_count": max(handoff_total - handoff_ready_count, 0),
        "blocked_count": total_handoff_blocked,
        "progress_pct": handoff_progress_pct,
        "detail": (
            "Las transiciones Solicitud a Orden, Orden a Recepción y Recepción a Cierre ya no tienen fricción documental."
            if handoff_ready_count == handoff_total and handoff_total
            else "Aún hay handoffs abiertos o bloqueados entre solicitudes, órdenes y recepciones."
        ),
    }
    closure_focus_rows: list[dict[str, Any]] = []
    handoff_focus_rows: list[dict[str, Any]] = []

    def _build_document_focus_rows(
        rows: list[dict[str, Any]],
        scope: str,
        *,
        fallback_key: str = "",
        fallback_label: str = "",
        fallback_count: int = 0,
        fallback_detail: str = "",
        fallback_action_label: str = "",
        fallback_action_url: str = "",
    ) -> list[dict[str, Any]]:
        scoped_rows = [row for row in rows if row.get("scope") == scope][:4]
        if scoped_rows:
            return [
                {
                    "scope": row.get("scope", scope),
                    "owner": row.get("owner") or _document_scope_owner(row.get("scope", scope)),
                    "folio": row.get("folio", ""),
                    "status": row.get("status", ""),
                    "detail": row.get("detail", ""),
                    "action_label": row.get("action_label", "Abrir"),
                    "action_url": row.get("action_url", ""),
                }
                for row in scoped_rows
            ]
        if fallback_count > 0:
            return [
                {
                    "scope": fallback_label or scope,
                    "owner": _document_scope_owner(scope),
                    "folio": fallback_key or "Resumen",
                    "status": f"{fallback_count} abiertos",
                    "detail": fallback_detail,
                    "action_label": fallback_action_label or "Abrir",
                    "action_url": fallback_action_url or "#",
                }
            ]
        return []

    if handoff_key != "auto":
        handoff_focus = next((item for item in handoff_checks if item["key"] == handoff_key), None)
    elif recepciones_total > 0:
        preferred_handoff_index = 2
        handoff_focus = (
            handoff_checks[preferred_handoff_index]
            if not handoff_checks[preferred_handoff_index]["is_ready"]
            else next((item for item in handoff_checks if not item["is_ready"]), None)
        )
    elif ordenes_total > 0:
        preferred_handoff_index = 1
        handoff_focus = (
            handoff_checks[preferred_handoff_index]
            if not handoff_checks[preferred_handoff_index]["is_ready"]
            else next((item for item in handoff_checks if not item["is_ready"]), None)
        )
    else:
        preferred_handoff_index = 0
        handoff_focus = (
            handoff_checks[preferred_handoff_index]
            if not handoff_checks[preferred_handoff_index]["is_ready"]
            else next((item for item in handoff_checks if not item["is_ready"]), None)
        )
    if handoff_focus:
        handoff_focus = {
            **handoff_focus,
            "tone": "danger" if int(handoff_focus["blocked_count"]) > 0 else "warning",
            "summary": f"La entrega entre etapas sigue abierta por: {handoff_focus['label'].lower()}.",
        }
        if handoff_focus["key"] == "solicitud_orden":
            handoff_focus_rows = _build_document_focus_rows(
                document_blocker_rows,
                "Solicitud",
                fallback_key="SOLICITUD-ORDEN",
                fallback_label="Solicitud → Orden",
                fallback_count=max(solicitudes_aprobadas - ordenes_total, 0),
                fallback_detail="Existen solicitudes aprobadas del plan que todavía no tienen orden emitida.",
                fallback_action_label=handoff_focus.get("action_label", "Abrir solicitudes"),
                fallback_action_url=handoff_focus.get("action_url", solicitudes_url),
            )
        elif handoff_focus["key"] == "orden_recepcion":
            handoff_focus_rows = _build_document_focus_rows(
                document_blocker_rows,
                "Orden",
                fallback_key="ORDEN-RECEPCION",
                fallback_label="Orden → Recepción",
                fallback_count=max((ordenes_confirmadas + ordenes_parciales + ordenes_cerradas) - recepciones_total, 0),
                fallback_detail="Existen órdenes confirmadas o parciales sin recepción asociada.",
                fallback_action_label=handoff_focus.get("action_label", "Abrir órdenes"),
                fallback_action_url=handoff_focus.get("action_url", ordenes_url),
            )
        elif handoff_focus["key"] == "recepcion_cierre":
            handoff_focus_rows = _build_document_focus_rows(
                document_blocker_rows,
                "Recepción",
                fallback_key="RECEPCION-CIERRE",
                fallback_label="Recepción → Cierre",
                fallback_count=recepciones_abiertas_total,
                fallback_detail="Existen recepciones pendientes o con diferencias que todavía no cierran el flujo.",
                fallback_action_label=handoff_focus.get("action_label", "Abrir recepciones"),
                fallback_action_url=handoff_focus.get("action_url", recepciones_url),
            )
    else:
        handoff_focus = {
            "label": "Handoffs completos",
            "is_ready": True,
            "ready_count": 0,
            "blocked_count": 0,
            "detail": "Las entregas entre solicitudes, órdenes y recepciones ya quedaron completas.",
            "action_label": "Abrir recepciones",
            "action_url": recepciones_url,
            "action_detail": "No quedan handoffs documentales abiertos para este plan.",
            "tone": "success",
            "summary": "El flujo entre etapas ya quedó completo para este plan.",
        }
    if closure_key != "auto":
        closure_focus = next((item for item in closure_checks if item["key"] == closure_key), None)
    elif recepciones_total > 0:
        preferred_closure_index = 2
        closure_focus = (
            closure_checks[preferred_closure_index]
            if not closure_checks[preferred_closure_index]["is_ready"]
            else next((item for item in closure_checks if not item["is_ready"]), None)
        )
    elif ordenes_total > 0:
        preferred_closure_index = 1
        closure_focus = (
            closure_checks[preferred_closure_index]
            if not closure_checks[preferred_closure_index]["is_ready"]
            else next((item for item in closure_checks if not item["is_ready"]), None)
        )
    else:
        preferred_closure_index = 0
        closure_focus = (
            closure_checks[preferred_closure_index]
            if not closure_checks[preferred_closure_index]["is_ready"]
            else next((item for item in closure_checks if not item["is_ready"]), None)
        )
    if closure_focus:
        closure_focus = {
            **closure_focus,
            "tone": "danger" if blocked_total > 0 else "warning",
            "summary": f"El plan sigue abierto por: {closure_focus['label'].lower()}.",
        }
        if closure_focus["key"] == "solicitudes_liberadas":
            closure_focus_rows = _build_document_focus_rows(
                document_blocker_rows,
                "Solicitud",
                fallback_key="CIERRE-SOLICITUDES",
                fallback_label="Solicitudes liberadas",
                fallback_count=solicitudes_pendientes_total,
                fallback_detail="Todavía hay solicitudes en borrador o revisión dentro del plan.",
                fallback_action_label=closure_focus.get("action_label", "Abrir solicitudes"),
                fallback_action_url=closure_focus.get("action_url", solicitudes_url),
            )
        elif closure_focus["key"] == "ordenes_sin_bloqueo":
            closure_focus_rows = _build_document_focus_rows(
                document_blocker_rows,
                "Orden",
                fallback_key="CIERRE-ORDENES",
                fallback_label="Órdenes sin bloqueo",
                fallback_count=ordenes_pendientes_total,
                fallback_detail="Todavía hay órdenes en borrador o enviadas sin cierre documental.",
                fallback_action_label=closure_focus.get("action_label", "Abrir órdenes"),
                fallback_action_url=closure_focus.get("action_url", ordenes_url),
            )
        elif closure_focus["key"] == "recepciones_aplicadas":
            closure_focus_rows = _build_document_focus_rows(
                document_blocker_rows,
                "Recepción",
                fallback_key="CIERRE-RECEPCIONES",
                fallback_label="Recepciones aplicadas",
                fallback_count=recepciones_abiertas_total,
                fallback_detail="Todavía hay recepciones abiertas que no han aplicado inventario o tienen diferencias.",
                fallback_action_label=closure_focus.get("action_label", "Abrir recepciones"),
                fallback_action_url=closure_focus.get("action_url", recepciones_url),
            )
    else:
        closure_focus = {
            "label": "Cierre documental completo",
            "is_ready": True,
            "detail": "Solicitudes, órdenes y recepciones ya cumplen el ciclo del plan.",
            "action_label": "Abrir recepciones",
            "action_url": recepciones_url,
            "action_detail": "No quedan bloqueos documentales abiertos para este plan.",
            "tone": "success",
            "summary": "El plan ya cerró criterios documentales en compras.",
        }

    valid_master_focus_keys = {
        str(card.get("key") or "").strip().lower() for card in master_blocker_class_cards if card.get("key")
    }
    valid_master_missing_keys = {
        str(card.get("key") or "").strip().lower() for card in master_blocker_missing_cards if card.get("key")
    }
    selected_master_focus_key = master_focus_key if master_focus_key in valid_master_focus_keys else "auto"
    selected_master_missing_key = master_missing_key if master_missing_key in valid_master_missing_keys else "auto"
    filtered_master_rows = (
        [
            row
            for row in master_blocker_detail_rows
            if str(row.get("class_key") or "").strip().lower() == selected_master_focus_key
        ]
        if selected_master_focus_key != "auto"
        else list(master_blocker_detail_rows)
    )
    if selected_master_missing_key != "auto":
        filtered_master_rows = [
            row
            for row in filtered_master_rows
            if str(row.get("missing_field") or "").strip().lower() == selected_master_missing_key
        ]

    master_focus_rows: list[dict[str, Any]] = []
    for row in filtered_master_rows[:3]:
        master_focus_rows.append(
            {
                "class_key": row.get("class_key", ""),
                "class_label": row.get("class_label", ""),
                "owner": row.get("owner") or "Maestros / DG",
                "missing_key": row.get("missing_field", ""),
                "insumo_nombre": row.get("name", ""),
                "missing_field": row.get("missing", ""),
                "detail": row.get("detail", ""),
                "action_label": row.get("action_label", "Abrir maestro"),
                "action_url": row.get("action_url", reverse("maestros:insumo_list")),
                "edit_url": row.get("edit_url", ""),
                "action_detail": row.get("action_detail", ""),
                "tone": "warning",
            }
        )
    if master_focus_rows:
        first_master_focus = master_focus_rows[0]
        master_focus = {
            **first_master_focus,
            "label": f"{first_master_focus['class_label']} · {first_master_focus['missing_field']}",
            "summary": (
                f"El flujo del plan sigue bloqueado por {first_master_focus['insumo_nombre']} "
                f"({first_master_focus['missing_field']})."
            ),
            "tone": "danger" if first_master_focus.get("tone") == "danger" else "warning",
        }
    else:
        master_focus = {
            "label": "Maestro ERP al día",
            "summary": "No se detectan bloqueos prioritarios del maestro para este plan.",
            "action_label": "Abrir maestro",
            "action_url": reverse("maestros:insumo_list"),
            "action_detail": "El catálogo no muestra bloqueos maestros sobre los documentos del plan.",
            "tone": "success",
            "class_label": "",
            "missing_field": "",
            "insumo_nombre": "",
            "count": 0,
        }

    return {
        "solicitudes_total": solicitudes_total,
        "solicitudes_borrador": solicitudes_borrador,
        "solicitudes_revision": solicitudes_revision,
        "solicitudes_aprobadas": solicitudes_aprobadas,
        "ordenes_total": ordenes_total,
        "ordenes_borrador": ordenes_borrador,
        "ordenes_enviadas": ordenes_enviadas,
        "ordenes_confirmadas": ordenes_confirmadas,
        "ordenes_parciales": ordenes_parciales,
        "ordenes_cerradas": ordenes_cerradas,
        "recepciones_total": recepciones_total,
        "recepciones_pendientes": recepciones_pendientes,
        "recepciones_diferencias": recepciones_diferencias,
        "recepciones_cerradas": recepciones_cerradas,
        "solicitudes_pendientes_total": solicitudes_pendientes_total,
        "ordenes_pendientes_total": ordenes_pendientes_total,
        "recepciones_abiertas_total": recepciones_abiertas_total,
        "blocked_total": blocked_total,
        "plan_status_label": plan_status_label,
        "plan_status_tone": plan_status_tone,
        "plan_status_detail": plan_status_detail,
        "consumo_aplicado": consumo_aplicado,
        "consumo_aplicado_en": consumo_aplicado_en,
        "consumo_aplicado_por": consumo_aplicado_por,
        "consumo_card_detail": consumo_card_detail,
        "consumo_card_tone": consumo_card_tone,
        "health_label": health_label,
        "health_tone": health_tone,
        "health_detail": health_detail,
        "stage_label": stage_label,
        "stage_tone": stage_tone,
        "stage_detail": stage_detail,
        "next_action_label": next_action_label,
        "next_action_url": next_action_url,
        "solicitudes_url": solicitudes_url,
        "ordenes_url": ordenes_url,
        "recepciones_url": recepciones_url,
        "document_cards": document_cards,
        "document_health_cards": document_health_cards,
        "document_stage_rows": document_stage_rows,
        "document_blocker_rows": document_blocker_rows[:12],
        "pipeline_steps": pipeline_steps,
        "purchase_gate": purchase_gate,
        "stage_focus": stage_focus,
        "selected_stage_key": stage_key,
        "closure_checks": closure_checks,
        "closure_summary": closure_summary,
        "closure_focus": closure_focus,
        "closure_focus_rows": closure_focus_rows,
        "selected_closure_key": closure_key,
        "handoff_checks": handoff_checks,
        "handoff_summary": handoff_summary,
        "handoff_focus": handoff_focus,
        "handoff_focus_rows": handoff_focus_rows,
        "selected_handoff_key": handoff_key,
        "master_blocker_class_cards": master_blocker_class_cards[:6],
        "master_blocker_missing_cards": master_blocker_missing_cards[:6],
        "master_blocker_detail_rows": filtered_master_rows[:12],
        "master_blocker_total": sum(int(card.get("count") or 0) for card in master_blocker_class_cards),
        "selected_master_focus_key": selected_master_focus_key,
        "selected_master_missing_key": selected_master_missing_key,
        "master_focus": master_focus,
        "master_focus_rows": master_focus_rows,
    }


def _export_plan_csv(plan: PlanProduccion, explosion: Dict[str, Any]) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    filename = f"plan_produccion_{plan.id}_{plan.fecha_produccion}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)

    writer.writerow(["PLAN DE PRODUCCION", plan.nombre])
    writer.writerow(["Fecha", plan.fecha_produccion.isoformat()])
    writer.writerow(["Estado", plan.get_estado_display()])
    writer.writerow(["Consumo aplicado", "SI" if plan.consumo_aplicado else "NO"])
    writer.writerow(["Consumo aplicado en", timezone.localtime(plan.consumo_aplicado_en).isoformat() if plan.consumo_aplicado_en else ""])
    writer.writerow(["Consumo aplicado por", str(plan.consumo_aplicado_por or "")])
    writer.writerow(["Cerrado en", timezone.localtime(plan.cerrado_en).isoformat() if plan.cerrado_en else ""])
    writer.writerow(["Cerrado por", str(plan.cerrado_por or "")])
    writer.writerow(["Costo total estimado", str(explosion["costo_total"])])
    writer.writerow([])
    writer.writerow(["PRODUCTOS EN PLAN"])
    writer.writerow(["Receta", "Cantidad", "Notas", "Costo estimado"])
    for row in explosion["items_detalle"]:
        writer.writerow(
            [
                row["receta"].nombre,
                f"{Decimal(str(row['cantidad'])):.3f}",
                row["notas"] or "",
                f"{Decimal(str(row['costo_estimado'])):.2f}",
            ]
        )

    writer.writerow([])
    writer.writerow(["INSUMOS CONSOLIDADOS"])
    writer.writerow(["Insumo", "Origen", "Proveedor sugerido", "Cantidad requerida", "Unidad", "Costo unitario", "Costo total"])
    for row in explosion["insumos"]:
        writer.writerow(
            [
                row["nombre"],
                row["origen"],
                row.get("proveedor_sugerido") or "-",
                f"{Decimal(str(row['cantidad'])):.3f}",
                row["unidad"],
                f"{Decimal(str(row['costo_unitario'])):.2f}",
                f"{Decimal(str(row['costo_total'])):.2f}",
            ]
        )

    return response


def _export_plan_xlsx(plan: PlanProduccion, explosion: Dict[str, Any]) -> HttpResponse:
    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen.append(["Plan", plan.nombre])
    ws_resumen.append(["Fecha", plan.fecha_produccion.isoformat()])
    ws_resumen.append(["Estado", plan.get_estado_display()])
    ws_resumen.append(["Consumo aplicado", "SI" if plan.consumo_aplicado else "NO"])
    ws_resumen.append(["Consumo aplicado en", timezone.localtime(plan.consumo_aplicado_en).isoformat() if plan.consumo_aplicado_en else ""])
    ws_resumen.append(["Consumo aplicado por", str(plan.consumo_aplicado_por or "")])
    ws_resumen.append(["Cerrado en", timezone.localtime(plan.cerrado_en).isoformat() if plan.cerrado_en else ""])
    ws_resumen.append(["Cerrado por", str(plan.cerrado_por or "")])
    ws_resumen.append(["Costo total estimado", float(explosion["costo_total"] or 0)])
    ws_resumen.append([])
    ws_resumen.append(["Productos", len(explosion["items_detalle"])])
    ws_resumen.append(["Insumos", len(explosion["insumos"])])

    ws_productos = wb.create_sheet("Productos")
    ws_productos.append(["Receta", "Cantidad", "Notas", "Costo estimado"])
    for row in explosion["items_detalle"]:
        ws_productos.append(
            [
                row["receta"].nombre,
                float(row["cantidad"] or 0),
                row["notas"] or "",
                float(row["costo_estimado"] or 0),
            ]
        )

    ws_insumos = wb.create_sheet("Insumos")
    ws_insumos.append(["Insumo", "Origen", "Proveedor sugerido", "Cantidad requerida", "Unidad", "Costo unitario", "Costo total"])
    for row in explosion["insumos"]:
        ws_insumos.append(
            [
                row["nombre"],
                row["origen"],
                row.get("proveedor_sugerido") or "-",
                float(row["cantidad"] or 0),
                row["unidad"],
                float(row["costo_unitario"] or 0),
                float(row["costo_total"] or 0),
            ]
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"plan_produccion_{plan.id}_{plan.fecha_produccion}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _export_plan_point_xlsx(plan: PlanProduccion) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Importacion Produccion"
    ws.append(["Código", "Nombre", "Cantidad Solicitada", "Is Insumo"])

    items = plan.items.select_related("receta").order_by("id")
    for item in items:
        receta = item.receta
        ws.append(
            [
                receta.codigo_point or "",
                receta.nombre,
                float(item.cantidad or 0),
                1 if receta.tipo == Receta.TIPO_PREPARACION else 0,
            ]
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"point_plan_produccion_{plan.id}_{plan.fecha_produccion}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _export_periodo_mrp_csv(resumen: Dict[str, Any]) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    filename = f"mrp_periodo_{resumen['periodo']}_{resumen['periodo_tipo']}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)

    writer.writerow(["MRP CONSOLIDADO POR PERIODO", resumen["periodo"]])
    writer.writerow(["Alcance", resumen["periodo_tipo"]])
    writer.writerow(["Planes", resumen["planes_count"]])
    writer.writerow(["Insumos", resumen["insumos_count"]])
    writer.writerow(["Costo total", str(resumen["costo_total"])])
    writer.writerow(["Alertas capacidad", resumen["alertas_capacidad"]])
    writer.writerow([])

    writer.writerow(["PLANES INCLUIDOS"])
    writer.writerow(["Plan", "Fecha producción", "Estado", "Consumo aplicado", "Cerrado", "Renglones"])
    for plan in resumen["planes"]:
        writer.writerow(
            [
                plan["nombre"],
                str(plan["fecha_produccion"]),
                plan.get("estado_label") or "",
                "SI" if plan.get("consumo_aplicado") else "NO",
                "SI" if plan.get("estado") == PlanProduccion.ESTADO_CERRADO else "NO",
                plan["items_count"],
            ]
        )
    writer.writerow([])

    writer.writerow(["INSUMOS CONSOLIDADOS"])
    writer.writerow(
        [
            "Insumo",
            "Origen",
            "Proveedor sugerido",
            "Cantidad requerida",
            "Stock actual",
            "Faltante",
            "Unidad",
            "Costo unitario",
            "Costo total",
            "Alerta capacidad",
        ]
    )
    for row in resumen["insumos"]:
        writer.writerow(
            [
                row["nombre"],
                row["origen"],
                row.get("proveedor_sugerido") or "-",
                f"{Decimal(str(row['cantidad'])):.3f}",
                f"{Decimal(str(row['stock_actual'])):.3f}",
                f"{Decimal(str(row['faltante'])):.3f}",
                row["unidad"],
                f"{Decimal(str(row['costo_unitario'])):.2f}",
                f"{Decimal(str(row['costo_total'])):.2f}",
                "SI" if row.get("alerta_capacidad") else "NO",
            ]
        )
    return response


def _export_periodo_mrp_xlsx(resumen: Dict[str, Any]) -> HttpResponse:
    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen.append(["Periodo", resumen["periodo"]])
    ws_resumen.append(["Alcance", resumen["periodo_tipo"]])
    ws_resumen.append(["Planes", resumen["planes_count"]])
    ws_resumen.append(["Insumos", resumen["insumos_count"]])
    ws_resumen.append(["Costo total", float(resumen["costo_total"] or 0)])
    ws_resumen.append(["Alertas capacidad", resumen["alertas_capacidad"]])

    ws_planes = wb.create_sheet("Planes")
    ws_planes.append(["Plan", "Fecha producción", "Estado", "Consumo aplicado", "Cerrado", "Renglones"])
    for plan in resumen["planes"]:
        ws_planes.append(
            [
                plan["nombre"],
                str(plan["fecha_produccion"]),
                plan.get("estado_label") or "",
                "SI" if plan.get("consumo_aplicado") else "NO",
                "SI" if plan.get("estado") == PlanProduccion.ESTADO_CERRADO else "NO",
                plan["items_count"],
            ]
        )

    ws_insumos = wb.create_sheet("Insumos")
    ws_insumos.append(
        [
            "Insumo",
            "Origen",
            "Proveedor sugerido",
            "Cantidad requerida",
            "Stock actual",
            "Faltante",
            "Unidad",
            "Costo unitario",
            "Costo total",
            "Alerta capacidad",
        ]
    )
    for row in resumen["insumos"]:
        ws_insumos.append(
            [
                row["nombre"],
                row["origen"],
                row.get("proveedor_sugerido") or "-",
                float(row["cantidad"] or 0),
                float(row["stock_actual"] or 0),
                float(row["faltante"] or 0),
                row["unidad"],
                float(row["costo_unitario"] or 0),
                float(row["costo_total"] or 0),
                "SI" if row.get("alerta_capacidad") else "NO",
            ]
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"mrp_periodo_{resumen['periodo']}_{resumen['periodo_tipo']}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _export_plan_status_dashboard_csv(summary: Dict[str, Any]) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="dg_estado_planes.csv"'
    writer = csv.writer(response)

    writer.writerow(["TABLERO DG ESTADO DE PLANES"])
    writer.writerow(["Estatus", summary["status"]])
    writer.writerow(["Detalle", summary["detail"]])
    writer.writerow(["Total", summary["total"]])
    writer.writerow(["Abiertos", summary["abiertos"]])
    writer.writerow(["Borrador", summary["borrador"]])
    writer.writerow(["Consumo aplicado", summary["consumo_aplicado"]])
    writer.writerow(["Cerrados", summary["cerrados"]])
    writer.writerow(["Rezago mas antiguo (dias)", summary["oldest_open_days"]])
    writer.writerow(["Plan mas antiguo abierto", summary["oldest_open_name"] or ""])
    writer.writerow(["Fecha mas antigua abierta", summary["oldest_open_date"] or ""])
    writer.writerow(["Fecha inicio", summary.get("start_date") or ""])
    writer.writerow(["Fecha fin", summary.get("end_date") or ""])
    writer.writerow(["Agrupacion", summary.get("group_by_label") or "Fecha producción"])
    writer.writerow([])
    writer.writerow([summary.get("group_by_label") or "Fecha producción", "Total", "Borrador", "Consumo aplicado", "Cerrado", "Abiertos"])
    for row in summary["rows"]:
        writer.writerow(
            [
                row["label"],
                row["total"],
                row["borrador"],
                row["consumo_aplicado"],
                row["cerrado"],
                row["abiertos"],
            ]
        )
    return response


def _export_plan_status_dashboard_xlsx(summary: Dict[str, Any]) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "DG Estados Plan"
    ws.append(["TABLERO DG ESTADO DE PLANES", ""])
    ws.append(["Estatus", summary["status"]])
    ws.append(["Detalle", summary["detail"]])
    ws.append(["Total", summary["total"]])
    ws.append(["Abiertos", summary["abiertos"]])
    ws.append(["Borrador", summary["borrador"]])
    ws.append(["Consumo aplicado", summary["consumo_aplicado"]])
    ws.append(["Cerrados", summary["cerrados"]])
    ws.append(["Rezago mas antiguo (dias)", summary["oldest_open_days"]])
    ws.append(["Plan mas antiguo abierto", summary["oldest_open_name"] or ""])
    ws.append(["Fecha mas antigua abierta", str(summary["oldest_open_date"] or "")])
    ws.append(["Fecha inicio", str(summary.get("start_date") or "")])
    ws.append(["Fecha fin", str(summary.get("end_date") or "")])
    ws.append(["Agrupacion", summary.get("group_by_label") or "Fecha producción"])
    ws.append([])
    ws.append([summary.get("group_by_label") or "Fecha producción", "Total", "Borrador", "Consumo aplicado", "Cerrado", "Abiertos"])
    for row in summary["rows"]:
        ws.append(
            [
                str(row["label"]),
                int(row["total"]),
                int(row["borrador"]),
                int(row["consumo_aplicado"]),
                int(row["cerrado"]),
                int(row["abiertos"]),
            ]
        )
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="dg_estado_planes.xlsx"'
    return response


def _normalize_periodo_mes(raw: str | None) -> str:
    txt = (raw or "").strip()
    if not txt:
        today = timezone.localdate()
        return f"{today.year:04d}-{today.month:02d}"

    txt = txt.replace("/", "-")
    if len(txt) >= 7:
        txt = txt[:7]
    try:
        y, m = txt.split("-")
        yi = int(y)
        mi = int(m)
        if 1 <= mi <= 12:
            return f"{yi:04d}-{mi:02d}"
    except Exception:
        pass

    today = timezone.localdate()
    return f"{today.year:04d}-{today.month:02d}"


def _document_scope_owner(scope: str) -> str:
    normalized = (scope or "").strip().lower()
    if normalized == "solicitud":
        return "Compras / Solicitante"
    if normalized == "orden":
        return "Compras"
    if normalized in {"recepción", "recepcion"}:
        return "Almacén / Recepción"
    if normalized == "receta":
        return "Producción / Costeo"
    if normalized == "insumo":
        return "Maestros / Compras"
    if normalized in {"forecast", "pronóstico", "pronostico"}:
        return "Ventas / Planeación"
    return "Operación ERP"


def _to_decimal_safe(value: object) -> Decimal:
    if value is None:
        return Decimal("0")
    try:
        txt = str(value).strip().replace(",", ".")
        if txt == "":
            return Decimal("0")
        return Decimal(txt)
    except Exception:
        return Decimal("0")


def _map_pronostico_header(header: str) -> str:
    key = normalizar_nombre(header).replace("_", " ")
    if key in {"receta", "producto", "nombre", "nombre receta"}:
        return "receta"
    if key in {"codigo point", "codigo", "sku"}:
        return "codigo_point"
    if key in {"periodo", "mes"}:
        return "periodo"
    if key in {"cantidad", "pronostico", "forecast"}:
        return "cantidad"
    return key


def _load_pronostico_rows(uploaded) -> list[dict]:
    filename = (uploaded.name or "").lower()
    rows: list[dict] = []
    if filename.endswith(".csv"):
        uploaded.seek(0)
        content = uploaded.read().decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(content.splitlines())
        for row in reader:
            parsed = {}
            for key, value in (row or {}).items():
                if not key:
                    continue
                parsed[_map_pronostico_header(str(key))] = value
            rows.append(parsed)
        return rows

    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        uploaded.seek(0)
        wb = load_workbook(uploaded, read_only=True, data_only=True)
        ws = wb.active
        values = list(ws.values)
        if not values:
            return []
        headers = [_map_pronostico_header(str(h or "")) for h in values[0]]
        for raw in values[1:]:
            parsed = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                parsed[header] = raw[idx] if idx < len(raw) else None
            rows.append(parsed)
        return rows

    raise ValueError("Formato no soportado. Usa CSV o XLSX.")


def _map_ventas_header(header: str) -> str:
    key = normalizar_nombre(header).replace("_", " ")
    if key in {"receta", "producto", "nombre receta", "nombre producto", "nombre"}:
        return "receta"
    if key in {"codigo point", "codigo", "sku", "codigo producto"}:
        return "codigo_point"
    if key in {"fecha", "dia", "date"}:
        return "fecha"
    if key in {"cantidad", "cantidad vendida", "unidades", "qty", "ventas"}:
        return "cantidad"
    if key in {"sucursal", "tienda", "store"}:
        return "sucursal"
    if key in {"sucursal codigo", "codigo sucursal", "store code"}:
        return "sucursal_codigo"
    if key in {"tickets", "ticket count", "num tickets"}:
        return "tickets"
    if key in {"monto", "total", "monto total", "importe"}:
        return "monto_total"
    return key


def _load_ventas_rows(uploaded) -> list[dict]:
    filename = (uploaded.name or "").lower()
    rows: list[dict] = []
    if filename.endswith(".csv"):
        uploaded.seek(0)
        content = uploaded.read().decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(content.splitlines())
        for row in reader:
            parsed = {}
            for key, value in (row or {}).items():
                if not key:
                    continue
                parsed[_map_ventas_header(str(key))] = value
            rows.append(parsed)
        return rows

    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        uploaded.seek(0)
        wb = load_workbook(uploaded, read_only=True, data_only=True)
        ws = wb.active
        values = list(ws.values)
        if not values:
            return []
        headers = [_map_ventas_header(str(h or "")) for h in values[0]]
        for raw in values[1:]:
            parsed = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                parsed[header] = raw[idx] if idx < len(raw) else None
            rows.append(parsed)
        return rows

    raise ValueError("Formato no soportado. Usa CSV o XLSX.")


def _map_solicitud_ventas_header(header: str) -> str:
    key = normalizar_nombre(header).replace("_", " ")
    if key in {"receta", "producto", "nombre receta", "nombre producto", "nombre"}:
        return "receta"
    if key in {"codigo point", "codigo", "sku", "codigo producto"}:
        return "codigo_point"
    if key in {"sucursal", "tienda", "store"}:
        return "sucursal"
    if key in {"sucursal codigo", "codigo sucursal", "store code"}:
        return "sucursal_codigo"
    if key in {"alcance", "tipo periodo", "scope"}:
        return "alcance"
    if key in {"periodo", "mes"}:
        return "periodo"
    if key in {"fecha", "fecha base", "base"}:
        return "fecha_base"
    if key in {"fecha inicio", "inicio", "start date"}:
        return "fecha_inicio"
    if key in {"fecha fin", "fin", "end date"}:
        return "fecha_fin"
    if key in {"cantidad", "cantidad solicitada", "qty", "solicitud"}:
        return "cantidad"
    return key


def _load_solicitud_ventas_rows(uploaded) -> list[dict]:
    filename = (uploaded.name or "").lower()
    rows: list[dict] = []
    if filename.endswith(".csv"):
        uploaded.seek(0)
        content = uploaded.read().decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(content.splitlines())
        for row in reader:
            parsed = {}
            for key, value in (row or {}).items():
                if not key:
                    continue
                parsed[_map_solicitud_ventas_header(str(key))] = value
            rows.append(parsed)
        return rows

    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        uploaded.seek(0)
        wb = load_workbook(uploaded, read_only=True, data_only=True)
        ws = wb.active
        values = list(ws.values)
        if not values:
            return []
        headers = [_map_solicitud_ventas_header(str(h or "")) for h in values[0]]
        for raw in values[1:]:
            parsed = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                parsed[header] = raw[idx] if idx < len(raw) else None
            rows.append(parsed)
        return rows

    raise ValueError("Formato no soportado. Usa CSV o XLSX.")


def _parse_date_safe(value: object) -> date | None:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    raw = raw.replace("/", "-")
    try:
        return date.fromisoformat(raw[:10])
    except Exception:
        pass
    parts = raw.split("-")
    if len(parts) == 3:
        try:
            if len(parts[0]) == 4:
                y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
            else:
                d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
            return date(y, m, d)
        except Exception:
            return None
    return None


def _to_int_safe(value: object, default: int = 0) -> int:
    if value is None:
        return default
    try:
        txt = str(value).strip()
        if txt == "":
            return default
        return int(float(txt))
    except Exception:
        return default


def _resolve_receta_for_sales(receta_name: str, codigo_point: str) -> Receta | None:
    receta = None
    if codigo_point:
        receta = Receta.objects.filter(codigo_point__iexact=codigo_point).order_by("id").first()
        if receta is None:
            codigo_norm = normalizar_codigo_point(codigo_point)
            if codigo_norm:
                alias = (
                    RecetaCodigoPointAlias.objects.filter(
                        codigo_point_normalizado=codigo_norm,
                        activo=True,
                    )
                    .select_related("receta")
                    .first()
                )
                if alias and alias.receta_id:
                    receta = alias.receta
    if receta is None and receta_name:
        receta = _resolve_receta_by_point_name_for_sales(receta_name)
    if receta is None and receta_name:
        receta = Receta.objects.filter(nombre_normalizado=normalizar_nombre(receta_name)).order_by("id").first()
        if receta is None:
            codigo_norm = normalizar_codigo_point(receta_name)
            if codigo_norm:
                receta = Receta.objects.filter(codigo_point__iexact=receta_name).order_by("id").first()
                if receta is None:
                    alias = (
                        RecetaCodigoPointAlias.objects.filter(
                            codigo_point_normalizado=codigo_norm,
                            activo=True,
                        )
                        .select_related("receta")
                        .first()
                    )
                    if alias and alias.receta_id:
                        receta = alias.receta
    return receta


def _resolve_receta_by_point_name_for_sales(receta_name: str) -> Receta | None:
    key = normalizar_nombre(receta_name)
    if not key:
        return None

    cache = getattr(_resolve_receta_by_point_name_for_sales, "_cache", None)
    if cache is None:
        cache = {}
        setattr(_resolve_receta_by_point_name_for_sales, "_cache", cache)
    if key in cache:
        return cache[key]

    receta = None
    alias_exact = (
        RecetaCodigoPointAlias.objects.filter(activo=True, nombre_point__iexact=(receta_name or "").strip())
        .select_related("receta")
        .order_by("id")
        .first()
    )
    if alias_exact and alias_exact.receta_id:
        receta = alias_exact.receta
    if receta is None:
        aliases = (
            RecetaCodigoPointAlias.objects.filter(activo=True)
            .exclude(nombre_point__isnull=True)
            .exclude(nombre_point="")
            .select_related("receta")
            .order_by("id")
        )
        for alias in aliases:
            if alias.receta_id and normalizar_nombre(alias.nombre_point or "") == key:
                receta = alias.receta
                break

    cache[key] = receta
    return receta


def _resolve_sucursal_for_sales(sucursal_name: str, sucursal_codigo: str, default_sucursal: Sucursal | None) -> Sucursal | None:
    sucursal = None
    if sucursal_codigo:
        sucursal = Sucursal.objects.filter(codigo__iexact=sucursal_codigo, activa=True).order_by("id").first()
    if sucursal is None and sucursal_name:
        sucursal = Sucursal.objects.filter(nombre__iexact=sucursal_name, activa=True).order_by("id").first()
    if sucursal is None and sucursal_name:
        objetivo = normalizar_nombre(sucursal_name)
        for row in sucursales_operativas().only("id", "codigo", "nombre").order_by("id"):
            if normalizar_nombre(row.nombre) == objetivo or normalizar_nombre(row.codigo) == objetivo:
                sucursal = row
                break
    return sucursal or default_sucursal


def _normalize_alcance_solicitud(raw: str | None) -> str:
    key = normalizar_nombre(str(raw or "")).replace(" ", "_")
    if key in {"mes", "mensual"}:
        return SolicitudVenta.ALCANCE_MES
    if key in {"semana", "semanal", "week"}:
        return SolicitudVenta.ALCANCE_SEMANA
    if key in {"fin_semana", "fin_de_semana", "weekend"}:
        return SolicitudVenta.ALCANCE_FIN_SEMANA
    return SolicitudVenta.ALCANCE_MES


def _ui_to_model_alcance(raw: str | None) -> str:
    key = (raw or "").strip().lower()
    if key == "semana":
        return SolicitudVenta.ALCANCE_SEMANA
    if key == "fin_semana":
        return SolicitudVenta.ALCANCE_FIN_SEMANA
    return SolicitudVenta.ALCANCE_MES


def _resolve_solicitud_window(
    *,
    alcance: str,
    periodo_raw: str | None,
    fecha_base_raw: object,
    fecha_inicio_raw: object,
    fecha_fin_raw: object,
    periodo_default: str,
    fecha_base_default: date,
) -> tuple[str, date, date]:
    if alcance == SolicitudVenta.ALCANCE_MES:
        periodo = _normalize_periodo_mes(periodo_raw or periodo_default)
        fecha_inicio, fecha_fin = _month_start_end(periodo)
        return periodo, fecha_inicio, fecha_fin

    fecha_inicio = _parse_date_safe(fecha_inicio_raw)
    fecha_fin = _parse_date_safe(fecha_fin_raw)
    if fecha_inicio and fecha_fin and fecha_fin >= fecha_inicio:
        periodo = f"{fecha_inicio.year:04d}-{fecha_inicio.month:02d}"
        return periodo, fecha_inicio, fecha_fin

    base = _parse_date_safe(fecha_base_raw) or fecha_inicio or fecha_base_default
    if alcance == SolicitudVenta.ALCANCE_FIN_SEMANA:
        fecha_inicio, fecha_fin = _weekend_start_end(base)
    else:
        fecha_inicio, fecha_fin = _week_start_end(base)
    periodo = f"{fecha_inicio.year:04d}-{fecha_inicio.month:02d}"
    return periodo, fecha_inicio, fecha_fin


def _month_start_end(periodo: str) -> tuple[date, date]:
    y_txt, m_txt = periodo.split("-")
    y = int(y_txt)
    m = int(m_txt)
    start = date(y, m, 1)
    end = date(y, m, monthrange(y, m)[1])
    return start, end


def _week_start_end(base: date) -> tuple[date, date]:
    start = base - timedelta(days=base.weekday())
    end = start + timedelta(days=6)
    return start, end


def _weekend_start_end(base: date) -> tuple[date, date]:
    wd = base.weekday()
    if wd == 5:
        start = base
    elif wd == 6:
        start = base - timedelta(days=1)
    else:
        start = base + timedelta(days=(5 - wd))
    return start, start + timedelta(days=1)


def _weighted_avg_decimal(values: list[Decimal]) -> Decimal:
    clean = [Decimal(str(v)) for v in values if Decimal(str(v)) >= 0]
    if not clean:
        return Decimal("0")
    total_weight = Decimal("0")
    total = Decimal("0")
    for idx, value in enumerate(clean, start=1):
        weight = Decimal(str(idx))
        total += value * weight
        total_weight += weight
    if total_weight <= 0:
        return Decimal("0")
    return total / total_weight


def _stddev_decimal(values: list[Decimal]) -> Decimal:
    clean = [Decimal(str(v)) for v in values]
    if len(clean) <= 1:
        return Decimal("0")
    mean = sum(clean, Decimal("0")) / Decimal(str(len(clean)))
    variance = sum(((v - mean) ** 2 for v in clean), Decimal("0")) / Decimal(str(len(clean)))
    return Decimal(str(sqrt(float(variance))))


def _forecast_month_qty(
    day_rows: list[tuple[date, Decimal]], target_start: date
) -> tuple[Decimal, int, Decimal, Decimal, int]:
    if not day_rows:
        return Decimal("0"), 0, Decimal("0"), Decimal("0"), 0

    monthly_totals: dict[tuple[int, int], Decimal] = defaultdict(lambda: Decimal("0"))
    for d, qty in day_rows:
        monthly_totals[(d.year, d.month)] += qty
    ordered_keys = sorted(monthly_totals.keys())
    ordered_values = [monthly_totals[k] for k in ordered_keys]

    recent_values = ordered_values[-3:] if len(ordered_values) >= 3 else ordered_values
    recent_avg = sum(recent_values, Decimal("0")) / Decimal(str(len(recent_values) or 1))

    seasonal_values = [v for (y, m), v in monthly_totals.items() if m == target_start.month and (y, m) != (target_start.year, target_start.month)]
    seasonal_avg = (
        sum(seasonal_values, Decimal("0")) / Decimal(str(len(seasonal_values)))
        if seasonal_values
        else recent_avg
    )

    trend_values = ordered_values[-6:] if len(ordered_values) >= 6 else ordered_values
    trend_next = recent_avg
    if len(trend_values) >= 2:
        slope = (trend_values[-1] - trend_values[0]) / Decimal(str(len(trend_values) - 1))
        trend_next = max(trend_values[-1] + slope, Decimal("0"))

    weighted = (
        (recent_avg * Decimal("0.50"))
        + (seasonal_avg * Decimal("0.30"))
        + (trend_next * Decimal("0.20"))
    )
    variance_samples = ordered_values[-12:] if len(ordered_values) >= 12 else ordered_values
    dispersion = _stddev_decimal(variance_samples)
    confidence = min(Decimal("1"), Decimal(str(len(ordered_values))) / Decimal("12"))
    return weighted, len(day_rows), confidence, dispersion, len(variance_samples)


def _forecast_range_qty(
    day_rows: list[tuple[date, Decimal]], target_start: date, target_end: date
) -> tuple[Decimal, int, Decimal, Decimal, int]:
    if not day_rows:
        return Decimal("0"), 0, Decimal("0"), Decimal("0"), 0

    by_dow: dict[int, list[Decimal]] = defaultdict(list)
    lower_window = target_start - timedelta(days=84)
    for d, qty in day_rows:
        if d < target_start and d >= lower_window:
            by_dow[d.weekday()].append(qty)

    all_daily = [qty for _, qty in day_rows[-120:]]
    global_avg = sum(all_daily, Decimal("0")) / Decimal(str(len(all_daily) or 1))

    horizon_days: list[date] = []
    pointer = target_start
    while pointer <= target_end:
        horizon_days.append(pointer)
        pointer += timedelta(days=1)

    used_samples = 0
    pred_total = Decimal("0")
    variance_samples: list[Decimal] = []
    for d in horizon_days:
        samples = by_dow.get(d.weekday(), [])
        if samples:
            recent = samples[-8:]
            day_pred = _weighted_avg_decimal(recent)
            used_samples += len(recent)
            variance_samples.extend(recent)
        else:
            day_pred = global_avg
            variance_samples.append(global_avg)
        pred_total += day_pred

    confidence = Decimal("0")
    if horizon_days:
        confidence = min(
            Decimal("1"),
            Decimal(str(used_samples)) / Decimal(str(len(horizon_days) * 8)),
        )
    dispersion = _stddev_decimal(variance_samples)
    return pred_total, len(day_rows), confidence, dispersion, len(variance_samples)


def _forecast_totals_from_rows(rows: list[dict[str, Any]]) -> dict[str, Decimal | int]:
    return {
        "recetas_count": len(rows),
        "forecast_total": sum((r["forecast_qty"] for r in rows), Decimal("0")).quantize(Decimal("0.001")),
        "forecast_low_total": sum((r["forecast_low"] for r in rows), Decimal("0")).quantize(Decimal("0.001")),
        "forecast_high_total": sum((r["forecast_high"] for r in rows), Decimal("0")).quantize(Decimal("0.001")),
        "pronostico_total": sum((r["pronostico_actual"] for r in rows), Decimal("0")).quantize(Decimal("0.001")),
        "delta_total": sum((r["delta"] for r in rows), Decimal("0")).quantize(Decimal("0.001")),
    }


def _forecast_history_meta(qs, *, alcance: str, target_start: date, target_end: date) -> dict[str, Any]:
    history_dates = list(qs.values_list("fecha", flat=True).distinct())
    if not history_dates:
        return {
            "available": False,
            "first_date": None,
            "last_date": None,
            "days_observed": 0,
            "years_observed": 0,
            "comparable_years": 0,
            "months_observed": 0,
            "scope_label": "Sin histórico",
        }

    first_date = min(history_dates)
    last_date = max(history_dates)
    years_observed = len({d.year for d in history_dates})
    months_observed = len({(d.year, d.month) for d in history_dates})
    if alcance == "mes":
        comparable_years = len({d.year for d in history_dates if d.month == target_start.month})
    else:
        target_weekdays = {((target_start.weekday() + offset) % 7) for offset in range((target_end - target_start).days + 1)}
        comparable_years = len({d.year for d in history_dates if d.weekday() in target_weekdays})
    return {
        "available": True,
        "first_date": first_date,
        "last_date": last_date,
        "days_observed": len(history_dates),
        "years_observed": years_observed,
        "comparable_years": comparable_years,
        "months_observed": months_observed,
        "scope_label": f"{first_date.isoformat()} a {last_date.isoformat()}",
    }


def _filter_forecast_result_by_confianza(
    result: dict[str, Any], min_confianza_pct: Decimal
) -> tuple[dict[str, Any], int]:
    if min_confianza_pct <= 0:
        return result, 0

    threshold = Decimal(str(min_confianza_pct))
    original_rows = list(result.get("rows") or [])
    kept_rows = [r for r in original_rows if Decimal(str(r.get("confianza") or 0)) >= threshold]
    filtered = len(original_rows) - len(kept_rows)
    if filtered <= 0:
        return result, 0

    updated = dict(result)
    updated["rows"] = kept_rows
    updated["totals"] = _forecast_totals_from_rows(kept_rows)
    return updated, filtered


def _build_forecast_from_history(
    *,
    alcance: str,
    periodo: str,
    fecha_base: date | None,
    sucursal: Sucursal | None,
    incluir_preparaciones: bool,
    safety_pct: Decimal,
) -> dict[str, Any]:
    if alcance == "mes":
        target_start, target_end = _month_start_end(periodo)
    else:
        base = fecha_base or timezone.localdate()
        if alcance == "fin_semana":
            target_start, target_end = _weekend_start_end(base)
        else:
            target_start, target_end = _week_start_end(base)

    qs = VentaHistorica.objects.filter(fecha__lt=target_start).select_related("receta", "sucursal")
    if sucursal:
        qs = qs.filter(sucursal=sucursal)
    if not incluir_preparaciones:
        qs = qs.filter(receta__tipo=Receta.TIPO_PRODUCTO_FINAL)
    history_meta = _forecast_history_meta(
        qs,
        alcance=alcance,
        target_start=target_start,
        target_end=target_end,
    )

    grouped: dict[int, dict[str, Any]] = {}
    for row in qs.values("receta_id", "receta__nombre", "fecha").annotate(total=Sum("cantidad")).order_by("receta_id", "fecha"):
        rid = int(row["receta_id"])
        item = grouped.setdefault(
            rid,
            {
                "receta_id": rid,
                "receta": row["receta__nombre"],
                "days": [],
            },
        )
        item["days"].append((row["fecha"], Decimal(str(row["total"] or 0))))

    periodo_target = f"{target_start.year:04d}-{target_start.month:02d}"
    pron_map = {
        p.receta_id: Decimal(str(p.cantidad or 0))
        for p in PronosticoVenta.objects.filter(periodo=periodo_target)
    }

    rows = []
    safety_factor = Decimal("1") + (safety_pct / Decimal("100"))
    for rid, item in grouped.items():
        days = item["days"]
        if alcance == "mes":
            predicted, obs, confidence, dispersion, samples_count = _forecast_month_qty(days, target_start)
        else:
            predicted, obs, confidence, dispersion, samples_count = _forecast_range_qty(days, target_start, target_end)
        predicted = max(predicted * safety_factor, Decimal("0"))
        predicted = predicted.quantize(Decimal("0.001"))
        dispersion = max(dispersion * safety_factor, Decimal("0")).quantize(Decimal("0.001"))
        forecast_low = max(predicted - dispersion, Decimal("0")).quantize(Decimal("0.001"))
        forecast_high = max(predicted + dispersion, Decimal("0")).quantize(Decimal("0.001"))
        if predicted <= 0:
            continue

        current = pron_map.get(rid, Decimal("0"))
        delta = predicted - current
        threshold = max(Decimal("1.000"), current * Decimal("0.05"))
        if delta > threshold:
            recomendacion = "SUBIR"
        elif delta < (threshold * Decimal("-1")):
            recomendacion = "BAJAR"
        else:
            recomendacion = "MANTENER"

        rows.append(
            {
                "receta_id": rid,
                "receta": item["receta"],
                "forecast_qty": predicted,
                "forecast_low": forecast_low,
                "forecast_high": forecast_high,
                "desviacion": dispersion,
                "muestras": int(samples_count),
                "pronostico_actual": current.quantize(Decimal("0.001")),
                "delta": delta.quantize(Decimal("0.001")),
                "recomendacion": recomendacion,
                "observaciones": obs,
                "confianza": (confidence * Decimal("100")).quantize(Decimal("0.1")),
            }
        )

    rows.sort(key=lambda x: x["forecast_qty"], reverse=True)
    return {
        "alcance": alcance,
        "periodo": periodo_target,
        "target_start": target_start,
        "target_end": target_end,
        "sucursal_id": sucursal.id if sucursal else None,
        "sucursal_nombre": f"{sucursal.codigo} - {sucursal.nombre}" if sucursal else "Todas",
        "rows": rows,
        "totals": _forecast_totals_from_rows(rows),
        "history_meta": history_meta,
    }


def _forecast_session_payload(result: dict[str, Any], top_rows: int = 80) -> dict[str, Any]:
    rows = []
    for row in result["rows"][:top_rows]:
        rows.append(
            {
                "receta_id": int(row["receta_id"]),
                "receta": row["receta"],
                "forecast_qty": float(row["forecast_qty"]),
                "forecast_low": float(row.get("forecast_low", row["forecast_qty"])),
                "forecast_high": float(row.get("forecast_high", row["forecast_qty"])),
                "desviacion": float(row.get("desviacion", 0)),
                "muestras": int(row.get("muestras", 0)),
                "pronostico_actual": float(row["pronostico_actual"]),
                "delta": float(row["delta"]),
                "recomendacion": row["recomendacion"],
                "observaciones": int(row["observaciones"]),
                "confianza": float(row["confianza"]),
            }
        )
    payload = {
        "alcance": result["alcance"],
        "periodo": result["periodo"],
        "target_start": str(result["target_start"]),
        "target_end": str(result["target_end"]),
        "sucursal_id": int(result["sucursal_id"]) if result["sucursal_id"] else None,
        "sucursal_nombre": result["sucursal_nombre"],
        "rows": rows,
        "totals": {
            "recetas_count": int(result["totals"]["recetas_count"]),
            "forecast_total": float(result["totals"]["forecast_total"]),
            "forecast_low_total": float(result["totals"].get("forecast_low_total", result["totals"]["forecast_total"])),
            "forecast_high_total": float(result["totals"].get("forecast_high_total", result["totals"]["forecast_total"])),
            "pronostico_total": float(result["totals"]["pronostico_total"]),
            "delta_total": float(result["totals"]["delta_total"]),
        },
        "history_meta": {
            "available": bool((result.get("history_meta") or {}).get("available")),
            "first_date": str((result.get("history_meta") or {}).get("first_date") or ""),
            "last_date": str((result.get("history_meta") or {}).get("last_date") or ""),
            "days_observed": int((result.get("history_meta") or {}).get("days_observed") or 0),
            "years_observed": int((result.get("history_meta") or {}).get("years_observed") or 0),
            "comparable_years": int((result.get("history_meta") or {}).get("comparable_years") or 0),
            "months_observed": int((result.get("history_meta") or {}).get("months_observed") or 0),
            "scope_label": str((result.get("history_meta") or {}).get("scope_label") or "Sin histórico"),
        },
    }
    if result.get("min_confianza_pct") is not None:
        payload["min_confianza_pct"] = float(result.get("min_confianza_pct") or 0)
    if result.get("escenario"):
        payload["escenario"] = str(result.get("escenario"))
    return payload


def _forecast_preview_operational_summary(payload: dict[str, Any] | None) -> dict[str, Any] | None:
    if not payload:
        return None
    rows = list(payload.get("rows") or [])
    if not rows:
        return {
            "available": False,
            "status": "Sin forecast",
            "tone": "warning",
            "detail": "No hay recetas suficientes para construir una base estadística utilizable.",
            "rows_count": 0,
            "avg_confidence": Decimal("0"),
            "with_history_count": 0,
            "top_rows": [],
            "scope_label": payload.get("sucursal_nombre") or "Todas",
        }

    avg_confidence = (
        sum((Decimal(str(row.get("confianza") or 0)) for row in rows), Decimal("0")) / Decimal(str(len(rows)))
    ).quantize(Decimal("0.1"))
    with_history_count = sum(1 for row in rows if int(row.get("observaciones") or 0) > 0)
    history_meta = payload.get("history_meta") or {}
    years_observed = int(history_meta.get("years_observed") or 0)
    comparable_years = int(history_meta.get("comparable_years") or 0)
    days_observed = int(history_meta.get("days_observed") or 0)
    if avg_confidence >= Decimal("70") and comparable_years >= 3:
        status = "Base robusta multianual"
        tone = "success"
        detail = "El forecast ya se apoya en temporadas comparables de varios años y sirve como base fuerte para plan y compras."
    elif avg_confidence >= Decimal("70"):
        status = "Base robusta"
        tone = "success"
        detail = "La base estadística ya es suficientemente estable para empujar ajuste y plan."
    elif avg_confidence >= Decimal("45") and years_observed >= 2:
        status = "Base utilizable multianual"
        tone = "warning"
        detail = "La base ya cubre más de un año, pero todavía requiere revisión comercial antes de tomarla como señal plena."
    elif avg_confidence >= Decimal("45"):
        status = "Base utilizable"
        tone = "warning"
        detail = "La base estadística sirve para orientar decisiones, pero todavía requiere revisión comercial."
    else:
        status = "Base limitada"
        tone = "danger"
        detail = "La cobertura histórica es débil y conviene tomar el forecast como referencia conservadora."

    top_rows = sorted(rows, key=lambda row: Decimal(str(row.get("forecast_qty") or 0)), reverse=True)[:5]
    return {
        "available": True,
        "status": status,
        "tone": tone,
        "detail": detail,
        "rows_count": len(rows),
        "avg_confidence": avg_confidence,
        "with_history_count": with_history_count,
        "top_rows": top_rows,
        "scope_label": payload.get("sucursal_nombre") or "Todas",
        "years_observed": years_observed,
        "comparable_years": comparable_years,
        "days_observed": days_observed,
        "history_span_label": history_meta.get("scope_label") or "Sin histórico",
    }


def _commercial_signal_gate(
    summary: dict[str, Any] | None,
    *,
    context_label: str,
    action_url: str,
    action_label: str,
) -> dict[str, Any]:
    if not summary or not bool(summary.get("available")):
        return {
            "status": "Sin base comercial",
            "tone": "danger",
            "is_ready": False,
            "blockers": 1,
            "detail": f"Todavía no hay una base comercial suficiente para sostener {context_label}.",
            "next_step": "Construir o revisar la base histórica antes de liberar la operación.",
            "action_url": action_url,
            "action_label": action_label,
        }

    tone = str(summary.get("tone") or "warning")
    years_observed = int(summary.get("years_observed") or 0)
    comparable_years = int(summary.get("comparable_years") or 0)
    status = str(summary.get("status") or "En revisión")
    detail = str(summary.get("detail") or "")

    if tone == "success":
        return {
            "status": status,
            "tone": "success",
            "is_ready": True,
            "blockers": 0,
            "detail": detail,
            "next_step": f"La base comercial ya puede respaldar {context_label}.",
            "action_url": action_url,
            "action_label": action_label,
        }
    if tone == "warning" and years_observed >= 2:
        return {
            "status": status,
            "tone": "warning",
            "is_ready": True,
            "blockers": 0,
            "detail": detail,
            "next_step": f"Opera {context_label} con criterio y valida el forecast antes de cerrar compras o producción.",
            "action_url": action_url,
            "action_label": action_label,
        }
    return {
        "status": status if comparable_years > 0 else "Base comercial frágil",
        "tone": "danger",
        "is_ready": False,
        "blockers": 1,
        "detail": detail or f"La señal comercial todavía es débil para sostener {context_label}.",
        "next_step": "Refuerza histórico, forecast o solicitud antes de cerrar el flujo troncal.",
        "action_url": action_url,
        "action_label": action_label,
    }


def _plan_master_demand_gate(plan: PlanProduccion | None, *, lookback_days: int = 60) -> dict[str, Any]:
    if not plan:
        return {
            "status": "Sin plan activo",
            "tone": "warning",
            "is_ready": False,
            "blockers": 0,
            "rows": [],
            "detail": "Todavía no existe un plan operativo para evaluar artículos maestros críticos por demanda.",
            "next_step": "Genera o selecciona un plan para revisar prioridades comerciales del maestro.",
            "action_url": reverse("recetas:plan_produccion"),
            "action_label": "Abrir plan",
        }

    plan_items = list(plan.items.select_related("receta").all())
    if not plan_items:
        return {
            "status": "Plan sin renglones",
            "tone": "warning",
            "is_ready": False,
            "blockers": 0,
            "rows": [],
            "detail": "El plan existe, pero todavía no tiene productos para evaluar prioridades del maestro.",
            "next_step": "Carga productos al plan antes de cerrar dependencias aguas abajo.",
            "action_url": reverse("recetas:plan_produccion") + f"?{urlencode({'plan_id': plan.id})}",
            "action_label": "Abrir plan",
        }

    plan_qty_map: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
    receta_ids: list[int] = []
    for item in plan_items:
        plan_qty_map[item.receta_id] += Decimal(str(item.cantidad or 0))
        receta_ids.append(item.receta_id)
    receta_ids = sorted(set(receta_ids))
    if not receta_ids:
        return {
            "status": "Plan sin productos válidos",
            "tone": "warning",
            "is_ready": False,
            "blockers": 0,
            "rows": [],
            "detail": "No hay productos válidos para medir impacto comercial del maestro.",
            "next_step": "Revisa la carga del plan y vuelve a evaluar dependencias.",
            "action_url": reverse("recetas:plan_produccion") + f"?{urlencode({'plan_id': plan.id})}",
            "action_label": "Abrir plan",
        }

    historico_map = {
        int(row["receta_id"]): Decimal(str(row["total"] or 0))
        for row in (
            VentaHistorica.objects.filter(
                receta_id__in=receta_ids,
                fecha__gte=plan.fecha_produccion - timedelta(days=lookback_days),
                fecha__lt=plan.fecha_produccion,
            )
            .values("receta_id")
            .annotate(total=Sum("cantidad"))
        )
    }
    priority_rows: dict[int, dict[str, Any]] = {}
    for linea in (
        LineaReceta.objects.filter(receta_id__in=receta_ids, insumo_id__isnull=False)
        .exclude(tipo_linea=LineaReceta.TIPO_SUBSECCION)
        .select_related("insumo", "receta")
    ):
        if not linea.insumo_id or not linea.insumo:
            continue
        canonical = canonical_insumo(linea.insumo) or linea.insumo
        readiness_profile = _insumo_erp_readiness(canonical)
        if readiness_profile["ready"]:
            continue
        row = priority_rows.setdefault(
            canonical.id,
            {
                "insumo_nombre": canonical.nombre,
                "historico_units": Decimal("0"),
                "required_qty": Decimal("0"),
                "missing": readiness_profile["missing"][:2] or ["Sin faltante"],
                "action_url": reverse("maestros:insumo_update", args=[canonical.id]),
            },
        )
        row["historico_units"] += historico_map.get(linea.receta_id, Decimal("0"))
        row["required_qty"] += plan_qty_map.get(linea.receta_id, Decimal("0")) * Decimal(str(linea.cantidad or 0))

    rows = sorted(
        priority_rows.values(),
        key=lambda item: (
            Decimal(str(item["historico_units"] or 0)),
            Decimal(str(item["required_qty"] or 0)),
        ),
        reverse=True,
    )
    top_rows = rows[:3]
    critical_rows = [row for row in rows if Decimal(str(row.get("historico_units") or 0)) >= Decimal("80")]
    if critical_rows:
        top_row = critical_rows[0]
        return {
            "status": "Demanda crítica bloqueada por maestro",
            "tone": "danger",
            "is_ready": False,
            "blockers": len(critical_rows),
            "rows": critical_rows[:3],
            "detail": (
                f"{len(critical_rows)} artículo(s) del plan ya sostienen demanda fuerte y siguen incompletos en maestro. "
                f"Prioridad actual: {top_row['insumo_nombre']}."
            ),
            "next_step": "Cierra primero el artículo maestro crítico antes de liberar BOM, MRP o compras.",
            "action_url": str(top_row["action_url"]),
            "action_label": "Cerrar prioridad crítica",
        }
    if rows:
        top_row = rows[0]
        return {
            "status": "Demanda priorizada en revisión",
            "tone": "warning",
            "is_ready": False,
            "blockers": len(rows),
            "rows": top_rows,
            "detail": "Hay artículos del plan con demanda relevante y faltantes del maestro que conviene cerrar antes de seguir aguas abajo.",
            "next_step": "Revisa primero los artículos del plan con mayor demanda antes de cerrar MRP o compras.",
            "action_url": str(top_row["action_url"]),
            "action_label": "Abrir artículos del plan",
        }
    return {
        "status": "Maestro comercial del plan estable",
        "tone": "success",
        "is_ready": True,
        "blockers": 0,
        "rows": [],
        "detail": "No hay artículos de alta demanda del plan bloqueados por faltantes del maestro.",
        "next_step": "Mantén seguimiento preventivo del maestro mientras avanzan plan y abastecimiento.",
        "action_url": reverse("maestros:insumo_list"),
        "action_label": "Abrir maestro",
    }


def _forecast_vs_solicitud_preview(payload: dict[str, Any] | None, escenario: str = "base") -> dict[str, Any] | None:
    if not payload:
        return None
    rows_payload = payload.get("rows") or []
    if not rows_payload:
        return None

    try:
        target_start = date.fromisoformat(str(payload.get("target_start")))
        target_end = date.fromisoformat(str(payload.get("target_end")))
    except Exception:
        return None

    escenario = (escenario or "base").strip().lower()
    if escenario not in {"base", "bajo", "alto"}:
        escenario = "base"

    sucursal_id = payload.get("sucursal_id")
    forecast_map: dict[int, dict[str, Any]] = {}
    for row in rows_payload:
        try:
            rid = int(row.get("receta_id") or 0)
        except Exception:
            rid = 0
        if rid <= 0:
            continue
        forecast_map[rid] = {
            "receta_id": rid,
            "receta": str(row.get("receta") or ""),
            "forecast_qty": Decimal(str(row.get("forecast_qty") or 0)),
            "forecast_low": Decimal(str(row.get("forecast_low") or row.get("forecast_qty") or 0)),
            "forecast_high": Decimal(str(row.get("forecast_high") or row.get("forecast_qty") or 0)),
        }

    if not forecast_map:
        return None

    solicitud_qs = SolicitudVenta.objects.filter(
        fecha_inicio__lte=target_end,
        fecha_fin__gte=target_start,
    )
    solicitud_qs = solicitud_qs.filter(alcance=_ui_to_model_alcance(str(payload.get("alcance") or "")))
    if sucursal_id:
        solicitud_qs = solicitud_qs.filter(sucursal_id=sucursal_id)
    solicitud_map = {
        int(r["receta_id"]): Decimal(str(r["total"] or 0))
        for r in solicitud_qs.values("receta_id").annotate(total=Sum("cantidad"))
    }

    missing_receta_ids = [rid for rid in solicitud_map.keys() if rid not in forecast_map]
    if missing_receta_ids:
        for receta in Receta.objects.filter(id__in=missing_receta_ids).only("id", "nombre"):
            forecast_map[receta.id] = {
                "receta_id": receta.id,
                "receta": receta.nombre,
                "forecast_qty": Decimal("0"),
                "forecast_low": Decimal("0"),
                "forecast_high": Decimal("0"),
            }

    rows: list[dict[str, Any]] = []
    for rid, base in forecast_map.items():
        forecast_qty_base = Decimal(str(base["forecast_qty"] or 0))
        forecast_low = Decimal(str(base.get("forecast_low") or forecast_qty_base or 0))
        forecast_high = Decimal(str(base.get("forecast_high") or forecast_qty_base or 0))
        if escenario == "bajo":
            forecast_qty = forecast_low
        elif escenario == "alto":
            forecast_qty = forecast_high
        else:
            forecast_qty = forecast_qty_base
        solicitud_qty = solicitud_map.get(rid, Decimal("0"))
        delta = solicitud_qty - forecast_qty
        tolerance = max(Decimal("1"), forecast_qty * Decimal("0.05"))
        if forecast_qty <= 0 and solicitud_qty > 0:
            status = "SIN_BASE"
            variacion_pct = None
        elif abs(delta) <= tolerance:
            status = "OK"
            variacion_pct = (
                (delta / forecast_qty) * Decimal("100")
                if forecast_qty > 0
                else Decimal("0")
            )
        elif delta > 0:
            status = "SOBRE"
            variacion_pct = (
                (delta / forecast_qty) * Decimal("100")
                if forecast_qty > 0
                else None
            )
        else:
            status = "BAJO"
            variacion_pct = (
                (delta / forecast_qty) * Decimal("100")
                if forecast_qty > 0
                else None
            )

        if forecast_qty <= 0 and solicitud_qty > 0:
            status_rango = "SIN_BASE"
        elif solicitud_qty < forecast_low:
            status_rango = "BAJO_RANGO"
        elif solicitud_qty > forecast_high:
            status_rango = "SOBRE_RANGO"
        else:
            status_rango = "EN_RANGO"

        rows.append(
            {
                "receta_id": rid,
                "receta": base["receta"],
                "forecast_qty": forecast_qty,
                "forecast_base": forecast_qty_base,
                "forecast_low": forecast_low,
                "forecast_high": forecast_high,
                "solicitud_qty": solicitud_qty,
                "delta_qty": delta,
                "variacion_pct": variacion_pct,
                "status": status,
                "status_rango": status_rango,
            }
        )

    rows.sort(key=lambda x: abs(x["delta_qty"]), reverse=True)
    total_forecast = sum((r["forecast_qty"] for r in rows), Decimal("0"))
    total_solicitud = sum((r["solicitud_qty"] for r in rows), Decimal("0"))
    return {
        "target_start": target_start,
        "target_end": target_end,
        "sucursal_id": sucursal_id,
        "sucursal_nombre": payload.get("sucursal_nombre") or "Todas",
        "escenario": escenario,
        "rows": rows,
        "totals": {
            "forecast_total": total_forecast,
            "solicitud_total": total_solicitud,
            "delta_total": total_solicitud - total_forecast,
            "ok_count": len([r for r in rows if r["status"] == "OK"]),
            "sobre_count": len([r for r in rows if r["status"] == "SOBRE"]),
            "bajo_count": len([r for r in rows if r["status"] == "BAJO"]),
            "sin_base_count": len([r for r in rows if r["status"] == "SIN_BASE"]),
            "en_rango_count": len([r for r in rows if r["status_rango"] == "EN_RANGO"]),
            "sobre_rango_count": len([r for r in rows if r["status_rango"] == "SOBRE_RANGO"]),
            "bajo_rango_count": len([r for r in rows if r["status_rango"] == "BAJO_RANGO"]),
        },
    }


def _forecast_vs_solicitud_operational_summary(compare: dict[str, Any] | None) -> dict[str, Any] | None:
    if not compare:
        return None
    totals = compare.get("totals") or {}
    row_count = len(compare.get("rows") or [])
    if row_count <= 0:
        return {
            "available": False,
            "status": "Sin comparativo",
            "tone": "warning",
            "detail": "No hay renglones suficientes para comparar forecast contra solicitud.",
            "alignment_pct": 0,
            "priority_label": "Sin datos",
            "priority_count": 0,
            "top_rows": [],
        }

    ok_count = int(totals.get("ok_count") or 0)
    alignment_pct = int(round((ok_count / row_count) * 100)) if row_count else 0
    sobre_count = int(totals.get("sobre_count") or 0)
    bajo_count = int(totals.get("bajo_count") or 0)
    sin_base_count = int(totals.get("sin_base_count") or 0)
    priority_label = "Sin base histórica"
    priority_count = sin_base_count
    tone = "warning"
    detail = "La sucursal necesita completar base histórica antes de confiar en la solicitud."
    if sobre_count >= bajo_count and sobre_count >= sin_base_count and sobre_count > 0:
        priority_label = "Sobre-solicitud"
        priority_count = sobre_count
        tone = "danger"
        detail = "La sucursal está pidiendo por arriba del forecast comparable y puede sobredemandar producción."
    elif bajo_count >= sobre_count and bajo_count >= sin_base_count and bajo_count > 0:
        priority_label = "Bajo-solicitud"
        priority_count = bajo_count
        tone = "warning"
        detail = "La sucursal está pidiendo por debajo del forecast comparable y puede quedarse corta."
    elif ok_count == row_count:
        priority_label = "Solicitud alineada"
        priority_count = ok_count
        tone = "success"
        detail = "La solicitud de ventas está completamente alineada con la base comparable."

    top_rows = sorted(
        list(compare.get("rows") or []),
        key=lambda row: abs(Decimal(str(row.get("delta_qty") or 0))),
        reverse=True,
    )[:5]
    return {
        "available": True,
        "status": "Alineación sólida" if alignment_pct >= 80 else "Alineación parcial" if alignment_pct >= 50 else "Alineación frágil",
        "tone": "success" if alignment_pct >= 80 else "warning" if alignment_pct >= 50 else "danger",
        "detail": detail,
        "alignment_pct": alignment_pct,
        "priority_label": priority_label,
        "priority_count": priority_count,
        "top_rows": top_rows,
    }


def _mrp_recipe_demand_signal(
    receta: Receta,
    forecast_preview: dict[str, Any] | None,
    forecast_vs_solicitud: dict[str, Any] | None,
    *,
    lookback_days: int = 60,
) -> dict[str, Any]:
    end_date = timezone.localdate() - timedelta(days=1)
    start_date = end_date - timedelta(days=max(lookback_days - 1, 0))
    history_qs = VentaHistorica.objects.filter(
        receta=receta,
        fecha__lt=timezone.localdate(),
    )
    history_meta = _forecast_history_meta(
        history_qs,
        alcance="semana",
        target_start=timezone.localdate(),
        target_end=timezone.localdate(),
    )
    rows_qs = VentaHistorica.objects.select_related("sucursal").filter(
        receta=receta,
        fecha__gte=start_date,
        fecha__lte=end_date,
    )
    days_count = rows_qs.values("fecha").distinct().count()
    branch_count = rows_qs.values("sucursal_id").distinct().count()
    total_units = rows_qs.aggregate(total=Sum("cantidad")).get("total") or Decimal("0")
    last_sale_date = rows_qs.order_by("-fecha").values_list("fecha", flat=True).first()
    top_branches = list(
        rows_qs.values("sucursal__codigo", "sucursal__nombre")
        .annotate(total=Sum("cantidad"))
        .order_by("-total", "sucursal__codigo")[:3]
    )
    avg_daily = (
        (Decimal(str(total_units)) / Decimal(str(days_count))).quantize(Decimal("0.1"))
        if days_count > 0
        else Decimal("0")
    )

    if receta.tipo != Receta.TIPO_PRODUCTO_FINAL and total_units <= 0:
        historico_status = "Sin venta directa"
        historico_tone = "warning"
        historico_detail = "Esta receta no vende directo; usa el MRP para validar estructura, costo y abastecimiento."
    elif days_count >= 12 and total_units > 0 and int(history_meta.get("comparable_years") or 0) >= 3:
        historico_status = "Base robusta multianual"
        historico_tone = "success"
        historico_detail = "La receta ya tiene señal reciente y temporadas comparables de varios años para orientar abastecimiento y capacidad."
    elif days_count >= 12 and total_units > 0:
        historico_status = "Base robusta"
        historico_tone = "success"
        historico_detail = "La receta ya tiene señal histórica suficiente para orientar abastecimiento y capacidad."
    elif days_count >= 5 and total_units > 0 and int(history_meta.get("years_observed") or 0) >= 2:
        historico_status = "Base utilizable multianual"
        historico_tone = "warning"
        historico_detail = "La receta ya tiene actividad reciente y más de un año de histórico, pero todavía conviene revisar el forecast antes de comprometer compra."
    elif days_count >= 5 and total_units > 0:
        historico_status = "Base utilizable"
        historico_tone = "warning"
        historico_detail = "La receta ya tiene actividad reciente, pero conviene revisar el forecast antes de comprometer compra."
    elif total_units > 0:
        historico_status = "Base limitada"
        historico_tone = "danger"
        historico_detail = "La receta tiene ventas recientes, pero todavía con poca cobertura para tomarla como señal fuerte."
    else:
        historico_status = "Sin histórico"
        historico_tone = "warning"
        historico_detail = "No hay ventas recientes para esta receta en la base histórica cargada."

    forecast_row = next(
        (
            row
            for row in (forecast_preview or {}).get("rows", [])
            if int(row.get("receta_id") or 0) == receta.id
        ),
        None,
    )
    if forecast_row:
        forecast_confidence = Decimal(str(forecast_row.get("confianza") or 0)).quantize(Decimal("0.1"))
        if forecast_confidence >= Decimal("70"):
            forecast_status = "Forecast confiable"
            forecast_tone = "success"
        elif forecast_confidence >= Decimal("45"):
            forecast_status = "Forecast utilizable"
            forecast_tone = "warning"
        else:
            forecast_status = "Forecast frágil"
            forecast_tone = "danger"
        forecast_detail = (
            f"Forecast {Decimal(str(forecast_row.get('forecast_qty') or 0)).quantize(Decimal('0.1'))} "
            f"con {forecast_confidence}% de confianza."
        )
    else:
        forecast_confidence = Decimal("0")
        forecast_status = "Sin forecast"
        forecast_tone = "warning"
        forecast_detail = "Todavía no hay forecast estadístico activo para esta receta en la sesión actual."

    compare_row = next(
        (
            row
            for row in (forecast_vs_solicitud or {}).get("rows", [])
            if int(row.get("receta_id") or 0) == receta.id
        ),
        None,
    )
    if compare_row:
        compare_status = str(compare_row.get("status") or "").upper()
        if compare_status == "OK":
            alignment_status = "Solicitud alineada"
            alignment_tone = "success"
            alignment_detail = "La solicitud comercial está dentro del rango razonable para esta receta."
        elif compare_status == "SOBRE":
            alignment_status = "Solicitud por arriba"
            alignment_tone = "danger"
            alignment_detail = "La solicitud supera el forecast comparable y puede sobredemandar producción o compra."
        elif compare_status == "BAJO":
            alignment_status = "Solicitud por debajo"
            alignment_tone = "warning"
            alignment_detail = "La solicitud está por debajo del forecast comparable y puede dejar corta la planeación."
        else:
            alignment_status = "Sin comparativo útil"
            alignment_tone = "warning"
            alignment_detail = "No hay base suficiente para contrastar la solicitud comercial contra el forecast."
    else:
        alignment_status = "Sin solicitud comparada"
        alignment_tone = "warning"
        alignment_detail = "No hay una solicitud comercial vigente para cruzar esta receta contra el forecast."

    return {
        "historico_status": historico_status,
        "historico_tone": historico_tone,
        "historico_detail": historico_detail,
        "historico_days": days_count,
        "historico_branches": branch_count,
        "historico_total": Decimal(str(total_units)).quantize(Decimal("0.1")),
        "historico_avg_daily": avg_daily,
        "last_sale_date": last_sale_date,
        "scope_label": f"{start_date.isoformat()} a {end_date.isoformat()}",
        "years_observed": int(history_meta.get("years_observed") or 0),
        "comparable_years": int(history_meta.get("comparable_years") or 0),
        "history_span_label": history_meta.get("scope_label") or "Sin histórico",
        "top_branches": top_branches,
        "forecast_status": forecast_status,
        "forecast_tone": forecast_tone,
        "forecast_detail": forecast_detail,
        "forecast_confidence": forecast_confidence,
        "forecast_qty": (
            Decimal(str(forecast_row.get("forecast_qty") or 0)).quantize(Decimal("0.1"))
            if forecast_row
            else Decimal("0")
        ),
        "alignment_status": alignment_status,
        "alignment_tone": alignment_tone,
        "alignment_detail": alignment_detail,
        "solicitud_qty": (
            Decimal(str(compare_row.get("solicitud_qty") or 0)).quantize(Decimal("0.1"))
            if compare_row
            else Decimal("0")
        ),
        "delta_qty": (
            Decimal(str(compare_row.get("delta_qty") or 0)).quantize(Decimal("0.1"))
            if compare_row
            else Decimal("0")
        ),
    }


def _forecast_backtest_windows(alcance: str, fecha_base: date, periods: int) -> list[tuple[date, date]]:
    windows: list[tuple[date, date]] = []
    if periods <= 0:
        return windows

    if alcance == "mes":
        anchor = date(fecha_base.year, fecha_base.month, 1)
        for i in range(periods):
            month = anchor.month - i
            year = anchor.year
            while month <= 0:
                month += 12
                year -= 1
            start = date(year, month, 1)
            end = date(year, month, monthrange(year, month)[1])
            windows.append((start, end))
        return windows

    if alcance == "fin_semana":
        current_start, current_end = _weekend_start_end(fecha_base)
    else:
        current_start, current_end = _week_start_end(fecha_base)

    for i in range(periods):
        delta_days = 7 * i
        windows.append((current_start - timedelta(days=delta_days), current_end - timedelta(days=delta_days)))
    return windows


def _build_forecast_backtest_preview(
    *,
    alcance: str,
    fecha_base: date,
    periods: int,
    sucursal: Sucursal | None,
    incluir_preparaciones: bool,
    safety_pct: Decimal,
    min_confianza_pct: Decimal,
    escenario: str,
    top: int,
) -> dict[str, Any] | None:
    windows = _forecast_backtest_windows(alcance, fecha_base, periods)
    if not windows:
        return None
    history_qs = VentaHistorica.objects.filter(fecha__lt=fecha_base)
    if sucursal:
        history_qs = history_qs.filter(sucursal=sucursal)
    if not incluir_preparaciones:
        history_qs = history_qs.filter(receta__tipo=Receta.TIPO_PRODUCTO_FINAL)
    history_meta = _forecast_history_meta(
        history_qs,
        alcance=alcance,
        target_start=fecha_base,
        target_end=fecha_base,
    )

    windows_payload: list[dict[str, Any]] = []
    sum_forecast_total = Decimal("0")
    sum_actual_total = Decimal("0")
    sum_abs_error = Decimal("0")
    ape_sum = Decimal("0")
    ape_count = 0

    for window_start, window_end in windows:
        periodo_window = f"{window_start.year:04d}-{window_start.month:02d}"
        forecast_result = _build_forecast_from_history(
            alcance=alcance,
            periodo=periodo_window,
            fecha_base=window_start,
            sucursal=sucursal,
            incluir_preparaciones=incluir_preparaciones,
            safety_pct=safety_pct,
        )
        forecast_result, _ = _filter_forecast_result_by_confianza(forecast_result, min_confianza_pct)
        qty_key = "forecast_qty"
        if escenario == "bajo":
            qty_key = "forecast_low"
        elif escenario == "alto":
            qty_key = "forecast_high"
        forecast_map = {
            int(row["receta_id"]): Decimal(str(row.get(qty_key) or 0))
            for row in (forecast_result.get("rows") or [])
        }

        actual_qs = VentaHistorica.objects.filter(fecha__gte=window_start, fecha__lte=window_end)
        if sucursal:
            actual_qs = actual_qs.filter(sucursal=sucursal)
        if not incluir_preparaciones:
            actual_qs = actual_qs.filter(receta__tipo=Receta.TIPO_PRODUCTO_FINAL)
        actual_map = {
            int(row["receta_id"]): Decimal(str(row["total"] or 0))
            for row in actual_qs.values("receta_id").annotate(total=Sum("cantidad"))
        }

        if min_confianza_pct > 0:
            union_ids = sorted(set(forecast_map.keys()))
        else:
            union_ids = sorted(set(forecast_map.keys()) | set(actual_map.keys()))
        if not union_ids:
            continue

        receta_names = {r.id: r.nombre for r in Receta.objects.filter(id__in=union_ids).only("id", "nombre")}
        rows: list[dict[str, Any]] = []
        forecast_total = Decimal("0")
        actual_total = Decimal("0")
        abs_error_total = Decimal("0")
        local_ape_sum = Decimal("0")
        local_ape_count = 0
        for receta_id in union_ids:
            forecast_qty = forecast_map.get(receta_id, Decimal("0"))
            actual_qty = actual_map.get(receta_id, Decimal("0"))
            delta_qty = forecast_qty - actual_qty
            abs_error = abs(delta_qty)

            forecast_total += forecast_qty
            actual_total += actual_qty
            abs_error_total += abs_error

            variacion_pct = None
            status_tag = "SIN_BASE"
            if actual_qty > 0:
                variacion_pct = ((delta_qty / actual_qty) * Decimal("100")).quantize(Decimal("0.1"))
                local_ape_sum += abs(variacion_pct)
                local_ape_count += 1
                if variacion_pct > Decimal("10"):
                    status_tag = "SOBRE"
                elif variacion_pct < Decimal("-10"):
                    status_tag = "BAJO"
                else:
                    status_tag = "OK"

            rows.append(
                {
                    "receta_id": receta_id,
                    "receta": receta_names.get(receta_id) or f"Receta {receta_id}",
                    "forecast_qty": float(forecast_qty),
                    "actual_qty": float(actual_qty),
                    "delta_qty": float(delta_qty),
                    "abs_error": float(abs_error),
                    "variacion_pct": float(variacion_pct) if variacion_pct is not None else None,
                    "status": status_tag,
                }
            )

        rows.sort(key=lambda r: abs(r["abs_error"]), reverse=True)
        mae = (abs_error_total / Decimal(str(len(union_ids)))).quantize(Decimal("0.001"))
        mape = None
        if local_ape_count > 0:
            mape = (local_ape_sum / Decimal(str(local_ape_count))).quantize(Decimal("0.1"))
            ape_sum += local_ape_sum
            ape_count += local_ape_count

        sum_forecast_total += forecast_total
        sum_actual_total += actual_total
        sum_abs_error += abs_error_total

        windows_payload.append(
            {
                "window_start": str(window_start),
                "window_end": str(window_end),
                "periodo": periodo_window,
                "recetas_count": len(union_ids),
                "forecast_total": float(forecast_total),
                "actual_total": float(actual_total),
                "bias_total": float((forecast_total - actual_total).quantize(Decimal("0.001"))),
                "mae": float(mae),
                "mape": float(mape) if mape is not None else None,
                "top_errors": rows[:top],
            }
        )

    if not windows_payload:
        return None

    overall_mape = None
    if ape_count > 0:
        overall_mape = (ape_sum / Decimal(str(ape_count))).quantize(Decimal("0.1"))
    overall_mae = (sum_abs_error / Decimal(str(max(1, len(windows_payload))))).quantize(Decimal("0.001"))
    return {
        "scope": {
            "alcance": alcance,
            "fecha_base": str(fecha_base),
            "periods": periods,
            "min_confianza_pct": float(min_confianza_pct),
            "escenario": escenario,
            "sucursal_id": sucursal.id if sucursal else None,
            "sucursal_nombre": f"{sucursal.codigo} - {sucursal.nombre}" if sucursal else "Todas",
        },
        "totals": {
            "windows_evaluated": len(windows_payload),
            "forecast_total": float(sum_forecast_total),
            "actual_total": float(sum_actual_total),
            "bias_total": float((sum_forecast_total - sum_actual_total).quantize(Decimal("0.001"))),
            "mae_promedio": float(overall_mae),
            "mape_promedio": float(overall_mape) if overall_mape is not None else None,
        },
        "windows": windows_payload,
        "history_meta": {
            "available": bool(history_meta.get("available")),
            "first_date": str(history_meta.get("first_date") or ""),
            "last_date": str(history_meta.get("last_date") or ""),
            "days_observed": int(history_meta.get("days_observed") or 0),
            "years_observed": int(history_meta.get("years_observed") or 0),
            "comparable_years": int(history_meta.get("comparable_years") or 0),
            "months_observed": int(history_meta.get("months_observed") or 0),
            "scope_label": str(history_meta.get("scope_label") or "Sin histórico"),
        },
    }


def _redirect_plan_produccion_with_request_params(request: HttpRequest) -> HttpResponse:
    next_params: dict[str, str] = {}
    for key in ("plan_id", "periodo", "forecast_compare_escenario"):
        value = (request.GET.get(key) or "").strip()
        if value:
            next_params[key] = value
    url = reverse("recetas:plan_produccion")
    if next_params:
        url += f"?{urlencode(next_params)}"
    return redirect(url)


def _forecast_vs_solicitud_filename(compare: dict[str, Any], export_format: str) -> str:
    start_txt = str(compare["target_start"]).replace("-", "")
    end_txt = str(compare["target_end"]).replace("-", "")
    escenario = str(compare.get("escenario") or "base").lower()
    return f"pronostico_vs_solicitud_{start_txt}_{end_txt}_{escenario}.{export_format}"


def _forecast_backtest_filename(payload: dict[str, Any], export_format: str) -> str:
    scope = payload.get("scope") or {}
    alcance = str(scope.get("alcance") or "mes").lower()
    fecha_base = str(scope.get("fecha_base") or timezone.localdate().isoformat()).replace("-", "")
    escenario = str(scope.get("escenario") or "base").lower()
    return f"forecast_backtest_{alcance}_{fecha_base}_{escenario}.{export_format}"


@login_required
@permission_required("recetas.view_planproduccion", raise_exception=True)
def forecast_preview_export(request: HttpRequest) -> HttpResponse:
    payload = request.session.get("pronostico_estadistico_preview")
    if not payload:
        messages.warning(request, "No hay preview de pronóstico para exportar.")
        return _redirect_plan_produccion_with_request_params(request)

    export_format = (request.GET.get("format") or "csv").strip().lower()
    start_txt = str(payload.get("target_start") or timezone.localdate().isoformat()).replace("-", "")
    end_txt = str(payload.get("target_end") or timezone.localdate().isoformat()).replace("-", "")
    alcance = str(payload.get("alcance") or "mes").lower()
    filename = f"forecast_preview_{alcance}_{start_txt}_{end_txt}.{'xlsx' if export_format == 'xlsx' else 'csv'}"
    rows = payload.get("rows") or []
    totals = payload.get("totals") or {}

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Alcance", str(payload.get("alcance") or "").upper()])
        ws_resumen.append(["Sucursal", payload.get("sucursal_nombre") or "Todas"])
        ws_resumen.append(["Periodo", payload.get("periodo") or ""])
        ws_resumen.append(["Rango inicio", payload.get("target_start") or ""])
        ws_resumen.append(["Rango fin", payload.get("target_end") or ""])
        ws_resumen.append(["Escenario", str(payload.get("escenario") or "base").upper()])
        ws_resumen.append(["Confianza minima %", float(payload.get("min_confianza_pct") or 0)])
        ws_resumen.append(["Recetas", int(totals.get("recetas_count") or 0)])
        ws_resumen.append(["Forecast total", float(totals.get("forecast_total") or 0)])
        ws_resumen.append(["Banda baja total", float(totals.get("forecast_low_total") or 0)])
        ws_resumen.append(["Banda alta total", float(totals.get("forecast_high_total") or 0)])
        ws_resumen.append(["Pronostico actual total", float(totals.get("pronostico_total") or 0)])
        ws_resumen.append(["Delta total", float(totals.get("delta_total") or 0)])

        ws_detalle = wb.create_sheet("Detalle")
        ws_detalle.append(
            [
                "Receta ID",
                "Receta",
                "Forecast",
                "Banda baja",
                "Banda alta",
                "Pronostico actual",
                "Delta",
                "Recomendacion",
                "Confianza %",
                "Desviacion",
                "Observaciones",
                "Muestras",
            ]
        )
        for row in rows:
            ws_detalle.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    float(row.get("forecast_qty") or 0),
                    float(row.get("forecast_low") or row.get("forecast_qty") or 0),
                    float(row.get("forecast_high") or row.get("forecast_qty") or 0),
                    float(row.get("pronostico_actual") or 0),
                    float(row.get("delta") or 0),
                    row.get("recomendacion") or "",
                    float(row.get("confianza") or 0),
                    float(row.get("desviacion") or 0),
                    row.get("observaciones") or "",
                    int(row.get("muestras") or 0),
                ]
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Alcance", str(payload.get("alcance") or "").upper()])
    writer.writerow(["Sucursal", payload.get("sucursal_nombre") or "Todas"])
    writer.writerow(["Periodo", payload.get("periodo") or ""])
    writer.writerow(["Rango inicio", payload.get("target_start") or ""])
    writer.writerow(["Rango fin", payload.get("target_end") or ""])
    writer.writerow(["Escenario", str(payload.get("escenario") or "base").upper()])
    writer.writerow(["Confianza minima %", f"{Decimal(str(payload.get('min_confianza_pct') or 0)):.1f}"])
    writer.writerow(["Recetas", int(totals.get("recetas_count") or 0)])
    writer.writerow(["Forecast total", f"{Decimal(str(totals.get('forecast_total') or 0)):.3f}"])
    writer.writerow(["Banda baja total", f"{Decimal(str(totals.get('forecast_low_total') or 0)):.3f}"])
    writer.writerow(["Banda alta total", f"{Decimal(str(totals.get('forecast_high_total') or 0)):.3f}"])
    writer.writerow(["Pronostico actual total", f"{Decimal(str(totals.get('pronostico_total') or 0)):.3f}"])
    writer.writerow(["Delta total", f"{Decimal(str(totals.get('delta_total') or 0)):.3f}"])
    writer.writerow([])
    writer.writerow(
        [
            "receta_id",
            "receta",
            "forecast",
            "banda_baja",
            "banda_alta",
            "pronostico_actual",
            "delta",
            "recomendacion",
            "confianza_pct",
            "desviacion",
            "observaciones",
            "muestras",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                f"{Decimal(str(row.get('forecast_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_low') or row.get('forecast_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_high') or row.get('forecast_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('pronostico_actual') or 0)):.3f}",
                f"{Decimal(str(row.get('delta') or 0)):.3f}",
                row.get("recomendacion") or "",
                f"{Decimal(str(row.get('confianza') or 0)):.1f}",
                f"{Decimal(str(row.get('desviacion') or 0)):.3f}",
                row.get("observaciones") or "",
                int(row.get("muestras") or 0),
            ]
        )
    return response


@login_required
@permission_required("recetas.view_planproduccion", raise_exception=True)
def forecast_vs_solicitud_export(request: HttpRequest) -> HttpResponse:
    payload = request.session.get("pronostico_estadistico_preview")
    escenario = (request.GET.get("escenario") or "base").strip().lower()
    if escenario not in {"base", "bajo", "alto"}:
        escenario = "base"
    compare = _forecast_vs_solicitud_preview(payload, escenario=escenario)
    if not compare:
        messages.warning(request, "No hay datos de Pronóstico vs Solicitud para exportar.")
        return _redirect_plan_produccion_with_request_params(request)

    export_format = (request.GET.get("format") or "csv").strip().lower()
    filename = _forecast_vs_solicitud_filename(compare, "xlsx" if export_format == "xlsx" else "csv")

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Sucursal", compare.get("sucursal_nombre") or "Todas"])
        ws_resumen.append(["Escenario", str(compare.get("escenario") or "base").upper()])
        ws_resumen.append(["Rango inicio", str(compare.get("target_start") or "")])
        ws_resumen.append(["Rango fin", str(compare.get("target_end") or "")])
        ws_resumen.append(["Forecast total", float(compare["totals"]["forecast_total"] or 0)])
        ws_resumen.append(["Solicitud total", float(compare["totals"]["solicitud_total"] or 0)])
        ws_resumen.append(["Delta total", float(compare["totals"]["delta_total"] or 0)])
        ws_resumen.append(["Alineadas", int(compare["totals"]["ok_count"] or 0)])
        ws_resumen.append(["Sobre solicitud", int(compare["totals"]["sobre_count"] or 0)])
        ws_resumen.append(["Bajo solicitud", int(compare["totals"]["bajo_count"] or 0)])
        ws_resumen.append(["Sin base", int(compare["totals"]["sin_base_count"] or 0)])

        ws_detalle = wb.create_sheet("Detalle")
        ws_detalle.append(
            [
                "Receta ID",
                "Receta",
                "Forecast",
                "Forecast base",
                "Forecast baja",
                "Forecast alta",
                "Solicitud",
                "Delta",
                "Variacion %",
                "Estatus",
                "Estatus rango",
            ]
        )
        status_labels = {"OK": "Alineada", "SOBRE": "Sobre", "BAJO": "Bajo", "SIN_BASE": "Sin base"}
        range_labels = {"EN_RANGO": "En rango", "SOBRE_RANGO": "Sobre rango", "BAJO_RANGO": "Bajo rango", "SIN_BASE": "Sin base"}
        for row in compare.get("rows") or []:
            variacion_pct = row.get("variacion_pct")
            ws_detalle.append(
                [
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    float(row.get("forecast_qty") or 0),
                    float(row.get("forecast_base") or 0),
                    float(row.get("forecast_low") or 0),
                    float(row.get("forecast_high") or 0),
                    float(row.get("solicitud_qty") or 0),
                    float(row.get("delta_qty") or 0),
                    float(variacion_pct) if variacion_pct is not None else None,
                    status_labels.get(str(row.get("status") or ""), str(row.get("status") or "-")),
                    range_labels.get(str(row.get("status_rango") or ""), str(row.get("status_rango") or "-")),
                ]
            )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Sucursal", compare.get("sucursal_nombre") or "Todas"])
    writer.writerow(["Escenario", str(compare.get("escenario") or "base").upper()])
    writer.writerow(["Rango inicio", str(compare.get("target_start") or "")])
    writer.writerow(["Rango fin", str(compare.get("target_end") or "")])
    writer.writerow(["Forecast total", f"{Decimal(str(compare['totals']['forecast_total'] or 0)):.3f}"])
    writer.writerow(["Solicitud total", f"{Decimal(str(compare['totals']['solicitud_total'] or 0)):.3f}"])
    writer.writerow(["Delta total", f"{Decimal(str(compare['totals']['delta_total'] or 0)):.3f}"])
    writer.writerow([])
    writer.writerow(
        [
            "receta_id",
            "receta",
            "forecast",
            "forecast_base",
            "forecast_baja",
            "forecast_alta",
            "solicitud",
            "delta",
            "variacion_pct",
            "estatus",
            "estatus_rango",
        ]
    )
    status_labels = {"OK": "Alineada", "SOBRE": "Sobre", "BAJO": "Bajo", "SIN_BASE": "Sin base"}
    range_labels = {"EN_RANGO": "En rango", "SOBRE_RANGO": "Sobre rango", "BAJO_RANGO": "Bajo rango", "SIN_BASE": "Sin base"}
    for row in compare.get("rows") or []:
        variacion_pct = row.get("variacion_pct")
        writer.writerow(
            [
                int(row.get("receta_id") or 0),
                row.get("receta") or "",
                f"{Decimal(str(row.get('forecast_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_base') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_low') or 0)):.3f}",
                f"{Decimal(str(row.get('forecast_high') or 0)):.3f}",
                f"{Decimal(str(row.get('solicitud_qty') or 0)):.3f}",
                f"{Decimal(str(row.get('delta_qty') or 0)):.3f}",
                f"{Decimal(str(variacion_pct)):.1f}" if variacion_pct is not None else "",
                status_labels.get(str(row.get("status") or ""), str(row.get("status") or "-")),
                range_labels.get(str(row.get("status_rango") or ""), str(row.get("status_rango") or "-")),
            ]
        )
    return response


@login_required
@permission_required("recetas.view_planproduccion", raise_exception=True)
def forecast_backtest_export(request: HttpRequest) -> HttpResponse:
    payload = request.session.get("pronostico_backtest_preview")
    if not payload:
        messages.warning(request, "No hay backtest generado para exportar.")
        return _redirect_plan_produccion_with_request_params(request)

    export_format = (request.GET.get("format") or "csv").strip().lower()
    filename = _forecast_backtest_filename(payload, "xlsx" if export_format == "xlsx" else "csv")
    scope = payload.get("scope") or {}
    totals = payload.get("totals") or {}
    windows = payload.get("windows") or []

    if export_format == "xlsx":
        wb = Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Alcance", str(scope.get("alcance") or "").upper()])
        ws_resumen.append(["Sucursal", scope.get("sucursal_nombre") or "Todas"])
        ws_resumen.append(["Escenario", str(scope.get("escenario") or "base").upper()])
        ws_resumen.append(["Fecha base", str(scope.get("fecha_base") or "")])
        ws_resumen.append(["Confianza minima %", float(scope.get("min_confianza_pct") or 0)])
        ws_resumen.append(["Ventanas evaluadas", int(totals.get("windows_evaluated") or 0)])
        ws_resumen.append(["Forecast total", float(totals.get("forecast_total") or 0)])
        ws_resumen.append(["Real total", float(totals.get("actual_total") or 0)])
        ws_resumen.append(["Bias total", float(totals.get("bias_total") or 0)])
        ws_resumen.append(["MAE promedio", float(totals.get("mae_promedio") or 0)])
        ws_resumen.append(["MAPE promedio", float(totals.get("mape_promedio")) if totals.get("mape_promedio") is not None else None])

        ws_windows = wb.create_sheet("Ventanas")
        ws_windows.append(["Inicio", "Fin", "Periodo", "Recetas", "Forecast", "Real", "Bias", "MAE", "MAPE"])
        for window in windows:
            ws_windows.append(
                [
                    window.get("window_start") or "",
                    window.get("window_end") or "",
                    window.get("periodo") or "",
                    int(window.get("recetas_count") or 0),
                    float(window.get("forecast_total") or 0),
                    float(window.get("actual_total") or 0),
                    float(window.get("bias_total") or 0),
                    float(window.get("mae") or 0),
                    float(window.get("mape")) if window.get("mape") is not None else None,
                ]
            )

        ws_errors = wb.create_sheet("TopErrores")
        ws_errors.append(
            [
                "Inicio",
                "Fin",
                "Periodo",
                "Receta ID",
                "Receta",
                "Forecast",
                "Real",
                "Delta",
                "Abs Error",
                "Variacion %",
                "Estatus",
            ]
        )
        for window in windows:
            for row in window.get("top_errors") or []:
                ws_errors.append(
                    [
                        window.get("window_start") or "",
                        window.get("window_end") or "",
                        window.get("periodo") or "",
                        int(row.get("receta_id") or 0),
                        row.get("receta") or "",
                        float(row.get("forecast_qty") or 0),
                        float(row.get("actual_qty") or 0),
                        float(row.get("delta_qty") or 0),
                        float(row.get("abs_error") or 0),
                        float(row.get("variacion_pct")) if row.get("variacion_pct") is not None else None,
                        row.get("status") or "",
                    ]
                )

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(["Alcance", str(scope.get("alcance") or "").upper()])
    writer.writerow(["Sucursal", scope.get("sucursal_nombre") or "Todas"])
    writer.writerow(["Escenario", str(scope.get("escenario") or "base").upper()])
    writer.writerow(["Fecha base", str(scope.get("fecha_base") or "")])
    writer.writerow(["Confianza minima %", f"{Decimal(str(scope.get('min_confianza_pct') or 0)):.1f}"])
    writer.writerow(["Ventanas evaluadas", int(totals.get("windows_evaluated") or 0)])
    writer.writerow(["Forecast total", f"{Decimal(str(totals.get('forecast_total') or 0)):.3f}"])
    writer.writerow(["Real total", f"{Decimal(str(totals.get('actual_total') or 0)):.3f}"])
    writer.writerow(["Bias total", f"{Decimal(str(totals.get('bias_total') or 0)):.3f}"])
    writer.writerow(["MAE promedio", f"{Decimal(str(totals.get('mae_promedio') or 0)):.3f}"])
    writer.writerow(
        [
            "MAPE promedio",
            f"{Decimal(str(totals.get('mape_promedio'))):.1f}" if totals.get("mape_promedio") is not None else "",
        ]
    )
    writer.writerow([])
    writer.writerow(["VENTANAS"])
    writer.writerow(["inicio", "fin", "periodo", "recetas", "forecast", "real", "bias", "mae", "mape"])
    for window in windows:
        writer.writerow(
            [
                window.get("window_start") or "",
                window.get("window_end") or "",
                window.get("periodo") or "",
                int(window.get("recetas_count") or 0),
                f"{Decimal(str(window.get('forecast_total') or 0)):.3f}",
                f"{Decimal(str(window.get('actual_total') or 0)):.3f}",
                f"{Decimal(str(window.get('bias_total') or 0)):.3f}",
                f"{Decimal(str(window.get('mae') or 0)):.3f}",
                f"{Decimal(str(window.get('mape'))):.1f}" if window.get("mape") is not None else "",
            ]
        )
    writer.writerow([])
    writer.writerow(["TOP_ERRORES"])
    writer.writerow(
        [
            "inicio",
            "fin",
            "periodo",
            "receta_id",
            "receta",
            "forecast",
            "real",
            "delta",
            "abs_error",
            "variacion_pct",
            "estatus",
        ]
    )
    for window in windows:
        for row in window.get("top_errors") or []:
            writer.writerow(
                [
                    window.get("window_start") or "",
                    window.get("window_end") or "",
                    window.get("periodo") or "",
                    int(row.get("receta_id") or 0),
                    row.get("receta") or "",
                    f"{Decimal(str(row.get('forecast_qty') or 0)):.3f}",
                    f"{Decimal(str(row.get('actual_qty') or 0)):.3f}",
                    f"{Decimal(str(row.get('delta_qty') or 0)):.3f}",
                    f"{Decimal(str(row.get('abs_error') or 0)):.3f}",
                    f"{Decimal(str(row.get('variacion_pct'))):.1f}" if row.get("variacion_pct") is not None else "",
                    row.get("status") or "",
                ]
            )
    return response


@login_required
@permission_required("recetas.change_planproduccion", raise_exception=True)
def pronosticos_descargar_plantilla(request: HttpRequest) -> HttpResponse:
    export_format = (request.GET.get("format") or "xlsx").strip().lower()
    headers = ["receta", "codigo_point", "periodo", "cantidad"]
    sample_rows = [
        ["Pastel Fresas Con Crema - Chico", "PFC-CHICO", timezone.localdate().strftime("%Y-%m"), "120"],
        ["Pan Vainilla Dawn - Chico", "PVD-CHICO", timezone.localdate().strftime("%Y-%m"), "220"],
    ]

    if export_format == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="plantilla_pronosticos.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(sample_rows)
        return response

    wb = Workbook()
    ws = wb.active
    ws.title = "pronosticos"
    ws.append(headers)
    for row in sample_rows:
        ws.append(row)
    for col in ("A", "B", "C", "D"):
        ws.column_dimensions[col].width = 28
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_pronosticos.xlsx"'
    return response


@login_required
@permission_required("recetas.change_planproduccion", raise_exception=True)
def ventas_historicas_descargar_plantilla(request: HttpRequest) -> HttpResponse:
    export_format = (request.GET.get("format") or "xlsx").strip().lower()
    headers = ["receta", "codigo_point", "sucursal_codigo", "sucursal", "fecha", "cantidad", "tickets", "monto_total"]
    sample_rows = [
        ["Pastel Fresas Con Crema - Chico", "PFC-CHICO", "MATRIZ", "Matriz", timezone.localdate().isoformat(), "24", "18", "3580"],
        ["Pastel Fresas Con Crema - Chico", "PFC-CHICO", "NORTE", "Sucursal Norte", timezone.localdate().isoformat(), "13", "9", "1970"],
    ]

    if export_format == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="plantilla_ventas_historicas.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(sample_rows)
        return response

    wb = Workbook()
    ws = wb.active
    ws.title = "ventas_historicas"
    ws.append(headers)
    for row in sample_rows:
        ws.append(row)
    for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 24
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_ventas_historicas.xlsx"'
    return response


@login_required
@permission_required("recetas.change_planproduccion", raise_exception=True)
def solicitud_ventas_descargar_plantilla(request: HttpRequest) -> HttpResponse:
    export_format = (request.GET.get("format") or "xlsx").strip().lower()
    headers = [
        "receta",
        "codigo_point",
        "sucursal_codigo",
        "sucursal",
        "alcance",
        "periodo",
        "fecha_inicio",
        "fecha_fin",
        "cantidad",
    ]
    sample_rows = [
        [
            "Pastel Fresas Con Crema - Chico",
            "PFC-CHICO",
            "MATRIZ",
            "Matriz",
            "MES",
            timezone.localdate().strftime("%Y-%m"),
            "",
            "",
            "120",
        ],
        [
            "Pastel Fresas Con Crema - Chico",
            "PFC-CHICO",
            "NORTE",
            "Sucursal Norte",
            "SEMANA",
            "",
            timezone.localdate().isoformat(),
            "",
            "34",
        ],
    ]

    if export_format == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="plantilla_solicitud_ventas.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(sample_rows)
        return response

    wb = Workbook()
    ws = wb.active
    ws.title = "solicitud_ventas"
    ws.append(headers)
    for row in sample_rows:
        ws.append(row)
    for col in ("A", "B", "C", "D", "E", "F", "G", "H", "I"):
        ws.column_dimensions[col].width = 24
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_solicitud_ventas.xlsx"'
    return response


@login_required
@permission_required("recetas.change_planproduccion", raise_exception=True)
@require_POST
def solicitud_ventas_guardar(request: HttpRequest) -> HttpResponse:
    plan_id = (request.POST.get("plan_id") or "").strip()
    next_params: dict[str, str] = {}
    if plan_id:
        next_params["plan_id"] = plan_id

    receta = Receta.objects.filter(pk=request.POST.get("receta_id")).first()
    if receta is None:
        messages.error(request, "Selecciona una receta válida para registrar solicitud de ventas.")
        if next_params:
            return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")
        return redirect(reverse("recetas:plan_produccion"))

    sucursal = Sucursal.objects.filter(pk=request.POST.get("sucursal_id")).first()
    alcance = _ui_to_model_alcance(request.POST.get("alcance"))
    periodo_default = _normalize_periodo_mes(request.POST.get("periodo"))
    fecha_base_default = _parse_date_safe(request.POST.get("fecha_base")) or timezone.localdate()
    cantidad = _to_decimal_safe(request.POST.get("cantidad"))
    if cantidad <= 0:
        messages.error(request, "La cantidad solicitada debe ser mayor a 0.")
        if next_params:
            return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")
        return redirect(reverse("recetas:plan_produccion"))

    periodo, fecha_inicio, fecha_fin = _resolve_solicitud_window(
        alcance=alcance,
        periodo_raw=request.POST.get("periodo"),
        fecha_base_raw=request.POST.get("fecha_base"),
        fecha_inicio_raw=request.POST.get("fecha_inicio"),
        fecha_fin_raw=request.POST.get("fecha_fin"),
        periodo_default=periodo_default,
        fecha_base_default=fecha_base_default,
    )
    fuente = (request.POST.get("fuente") or "UI_SOL_VENTAS").strip()[:40] or "UI_SOL_VENTAS"
    solicitud, created = SolicitudVenta.objects.get_or_create(
        receta=receta,
        sucursal=sucursal,
        alcance=alcance,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        defaults={
            "periodo": periodo,
            "cantidad": cantidad,
            "fuente": fuente,
        },
    )
    if created:
        messages.success(request, "Solicitud de ventas registrada.")
    else:
        solicitud.periodo = periodo
        solicitud.cantidad = cantidad
        solicitud.fuente = fuente
        solicitud.save(update_fields=["periodo", "cantidad", "fuente", "actualizado_en"])
        messages.success(request, "Solicitud de ventas actualizada.")

    next_params["periodo"] = periodo
    if next_params:
        return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")
    return redirect(reverse("recetas:plan_produccion"))


@login_required
@permission_required("recetas.change_planproduccion", raise_exception=True)
@require_POST
def solicitud_ventas_importar(request: HttpRequest) -> HttpResponse:
    uploaded = request.FILES.get("archivo")
    plan_id = (request.POST.get("plan_id") or "").strip()
    modo = (request.POST.get("modo") or "replace").strip().lower()
    if modo not in {"replace", "accumulate"}:
        modo = "replace"
    periodo_default = _normalize_periodo_mes(request.POST.get("periodo_default"))
    fecha_base_default = _parse_date_safe(request.POST.get("fecha_base_default")) or timezone.localdate()
    alcance_default = _ui_to_model_alcance(request.POST.get("alcance_default"))
    fuente = (request.POST.get("fuente") or "UI_SOL_VENTAS").strip()[:40] or "UI_SOL_VENTAS"
    sucursal_default = Sucursal.objects.filter(pk=request.POST.get("sucursal_default_id")).first()

    next_params: dict[str, str] = {}
    if plan_id:
        next_params["plan_id"] = plan_id
    next_params["periodo"] = periodo_default

    if not uploaded:
        messages.error(request, "Selecciona un archivo para importar solicitudes de ventas.")
        return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")

    try:
        rows = _load_solicitud_ventas_rows(uploaded)
    except Exception as exc:
        messages.error(request, f"No se pudo leer el archivo: {exc}")
        return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")

    if not rows:
        messages.warning(request, "Archivo sin filas válidas para solicitud de ventas.")
        return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")

    created = 0
    updated = 0
    skipped = 0
    unresolved_recetas: list[str] = []
    unresolved_sucursales: list[str] = []

    for row in rows:
        receta_name = str(row.get("receta") or "").strip()
        codigo_point = str(row.get("codigo_point") or "").strip()
        receta = _resolve_receta_for_sales(receta_name, codigo_point)
        if receta is None:
            unresolved_recetas.append(receta_name or codigo_point or "sin_identificador")
            skipped += 1
            continue

        cantidad = _to_decimal_safe(row.get("cantidad"))
        if cantidad < 0:
            skipped += 1
            continue

        sucursal_name = str(row.get("sucursal") or "").strip()
        sucursal_codigo = str(row.get("sucursal_codigo") or "").strip()
        sucursal = _resolve_sucursal_for_sales(sucursal_name, sucursal_codigo, sucursal_default)
        if (sucursal_name or sucursal_codigo) and sucursal is None:
            unresolved_sucursales.append(sucursal_codigo or sucursal_name)
            skipped += 1
            continue

        alcance = _normalize_alcance_solicitud(str(row.get("alcance") or ""))
        if not str(row.get("alcance") or "").strip():
            alcance = alcance_default
        periodo, fecha_inicio, fecha_fin = _resolve_solicitud_window(
            alcance=alcance,
            periodo_raw=str(row.get("periodo") or ""),
            fecha_base_raw=row.get("fecha_base"),
            fecha_inicio_raw=row.get("fecha_inicio"),
            fecha_fin_raw=row.get("fecha_fin"),
            periodo_default=periodo_default,
            fecha_base_default=fecha_base_default,
        )

        record = SolicitudVenta.objects.filter(
            receta=receta,
            sucursal=sucursal,
            alcance=alcance,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
        ).first()
        if record:
            if modo == "accumulate":
                record.cantidad = Decimal(str(record.cantidad or 0)) + cantidad
            else:
                record.cantidad = cantidad
            record.periodo = periodo
            record.fuente = fuente
            record.save(update_fields=["cantidad", "periodo", "fuente", "actualizado_en"])
            updated += 1
        else:
            SolicitudVenta.objects.create(
                receta=receta,
                sucursal=sucursal,
                alcance=alcance,
                periodo=periodo,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                cantidad=cantidad,
                fuente=fuente,
            )
            created += 1

    messages.success(
        request,
        f"Solicitudes de ventas importadas. Creadas: {created}. Actualizadas: {updated}. Omitidas: {skipped}.",
    )
    if unresolved_recetas:
        sample = ", ".join(unresolved_recetas[:5])
        extra = "" if len(unresolved_recetas) <= 5 else f" (+{len(unresolved_recetas) - 5} más)"
        messages.warning(request, f"Sin receta equivalente para: {sample}{extra}.")
    if unresolved_sucursales:
        sample = ", ".join(unresolved_sucursales[:5])
        extra = "" if len(unresolved_sucursales) <= 5 else f" (+{len(unresolved_sucursales) - 5} más)"
        messages.warning(request, f"Sucursales no encontradas: {sample}{extra}.")
    return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")


@login_required
@permission_required("recetas.change_planproduccion", raise_exception=True)
@require_POST
def solicitud_ventas_aplicar_desde_forecast(request: HttpRequest) -> HttpResponse:
    plan_id = (request.POST.get("plan_id") or "").strip()
    next_params: dict[str, str] = {}
    if plan_id:
        next_params["plan_id"] = plan_id

    payload = request.session.get("pronostico_estadistico_preview")
    if not payload:
        messages.error(request, "No hay preview de pronóstico activo. Ejecuta primero la previsualización.")
        if next_params:
            return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")
        return redirect(reverse("recetas:plan_produccion"))

    sucursal_id = payload.get("sucursal_id")
    if not sucursal_id:
        messages.error(request, "Selecciona una sucursal en la previsualización para aplicar ajuste automático.")
        if next_params:
            return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")
        return redirect(reverse("recetas:plan_produccion"))

    escenario = (request.POST.get("escenario") or payload.get("escenario") or "base").strip().lower()
    if escenario not in {"base", "bajo", "alto"}:
        escenario = "base"
    next_params["forecast_compare_escenario"] = escenario
    compare = _forecast_vs_solicitud_preview(payload, escenario=escenario)
    if not compare or not compare.get("rows"):
        messages.warning(request, "No hay filas disponibles para aplicar ajustes.")
        if next_params:
            return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")
        return redirect(reverse("recetas:plan_produccion"))

    modo = (request.POST.get("modo") or "desviadas").strip().lower()
    receta_id = _to_int_safe(request.POST.get("receta_id"), default=0)
    max_variacion_pct = None
    max_variacion_raw = (request.POST.get("max_variacion_pct") or "").strip()
    if max_variacion_raw:
        parsed_cap = _to_decimal_safe(max_variacion_raw)
        if parsed_cap < 0:
            parsed_cap = Decimal("0")
        max_variacion_pct = parsed_cap
    rows = list(compare["rows"])
    if modo == "sobre":
        rows = [r for r in rows if r["status"] == "SOBRE"]
    elif modo == "bajo":
        rows = [r for r in rows if r["status"] == "BAJO"]
    elif modo == "receta":
        rows = [r for r in rows if int(r["receta_id"]) == receta_id]
    else:
        rows = [r for r in rows if r["status"] in {"SOBRE", "BAJO"}]

    if not rows:
        messages.info(request, "No hay filas objetivo para el ajuste seleccionado.")
        if next_params:
            return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")
        return redirect(reverse("recetas:plan_produccion"))

    alcance = _ui_to_model_alcance(payload.get("alcance"))
    periodo = _normalize_periodo_mes(payload.get("periodo"))
    target_start = _parse_date_safe(payload.get("target_start")) or compare["target_start"]
    target_end = _parse_date_safe(payload.get("target_end")) or compare["target_end"]
    fuente = (request.POST.get("fuente") or "AUTO_FORECAST_ADJUST").strip()[:40] or "AUTO_FORECAST_ADJUST"

    sucursal = Sucursal.objects.filter(pk=sucursal_id).first()
    if sucursal is None:
        messages.error(request, "La sucursal del preview ya no existe; vuelve a generar el pronóstico.")
        if next_params:
            return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")
        return redirect(reverse("recetas:plan_produccion"))

    created = 0
    updated = 0
    skipped = 0
    skipped_cap = 0
    for row in rows:
        forecast_qty = Decimal(str(row.get("forecast_qty") or 0))
        if forecast_qty < 0:
            skipped += 1
            continue
        receta = Receta.objects.filter(pk=row["receta_id"]).first()
        if receta is None:
            skipped += 1
            continue
        record, was_created = SolicitudVenta.objects.get_or_create(
            receta=receta,
            sucursal=sucursal,
            alcance=alcance,
            fecha_inicio=target_start,
            fecha_fin=target_end,
            defaults={
                "periodo": periodo,
                "cantidad": forecast_qty,
                "fuente": fuente,
            },
        )
        if was_created:
            created += 1
            continue
        old_qty = Decimal(str(record.cantidad or 0))
        if max_variacion_pct is not None and old_qty > 0:
            variacion_pct = abs(((forecast_qty - old_qty) / old_qty) * Decimal("100"))
            if variacion_pct > max_variacion_pct:
                skipped += 1
                skipped_cap += 1
                continue
        record.periodo = periodo
        record.cantidad = forecast_qty
        record.fuente = fuente
        record.save(update_fields=["periodo", "cantidad", "fuente", "actualizado_en"])
        updated += 1

    if max_variacion_pct is None:
        messages.success(
            request,
            (
                f"Ajuste aplicado desde forecast ({escenario}). "
                f"Creadas: {created}. Actualizadas: {updated}. Omitidas: {skipped}."
            ),
        )
    else:
        messages.success(
            request,
            (
                f"Ajuste aplicado con tope {max_variacion_pct}% ({escenario})."
                f" Creadas: {created}. Actualizadas: {updated}. Omitidas: {skipped}."
                f" Omitidas por tope: {skipped_cap}."
            ),
        )
    next_params["periodo"] = periodo
    return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")


@login_required
@permission_required("recetas.change_planproduccion", raise_exception=True)
@require_POST
def ventas_historicas_importar(request: HttpRequest) -> HttpResponse:
    uploaded = request.FILES.get("archivo")
    plan_id = (request.POST.get("plan_id") or "").strip()
    modo = (request.POST.get("modo") or "replace").strip().lower()
    if modo not in {"replace", "accumulate"}:
        modo = "replace"
    fuente = (request.POST.get("fuente") or "UI_VENTAS").strip()[:40] or "UI_VENTAS"
    sucursal_default = Sucursal.objects.filter(pk=request.POST.get("sucursal_default_id")).first()

    next_url = reverse("recetas:plan_produccion")
    if plan_id:
        next_url = f"{next_url}?{urlencode({'plan_id': plan_id})}"

    if not uploaded:
        messages.error(request, "Selecciona un archivo para importar historial de ventas.")
        return redirect(next_url)

    try:
        rows = _load_ventas_rows(uploaded)
    except Exception as exc:
        messages.error(request, f"No se pudo leer el archivo: {exc}")
        return redirect(next_url)

    if not rows:
        messages.warning(request, "Archivo sin filas válidas para historial de ventas.")
        return redirect(next_url)

    created = 0
    updated = 0
    skipped = 0
    unresolved_recetas: list[str] = []
    unresolved_sucursales: list[str] = []

    for row in rows:
        receta_name = str(row.get("receta") or "").strip()
        codigo_point = str(row.get("codigo_point") or "").strip()
        receta = _resolve_receta_for_sales(receta_name, codigo_point)
        if receta is None:
            unresolved_recetas.append(receta_name or codigo_point or "sin_identificador")
            skipped += 1
            continue

        fecha = _parse_date_safe(row.get("fecha"))
        if not fecha:
            skipped += 1
            continue

        cantidad = _to_decimal_safe(row.get("cantidad"))
        if cantidad < 0:
            skipped += 1
            continue

        sucursal_name = str(row.get("sucursal") or "").strip()
        sucursal_codigo = str(row.get("sucursal_codigo") or "").strip()
        sucursal = _resolve_sucursal_for_sales(sucursal_name, sucursal_codigo, sucursal_default)
        if (sucursal_name or sucursal_codigo) and sucursal is None:
            unresolved_sucursales.append(sucursal_codigo or sucursal_name)
            skipped += 1
            continue

        tickets = max(0, _to_int_safe(row.get("tickets"), default=0))
        monto_total = _to_decimal_safe(row.get("monto_total"))

        existing_qs = VentaHistorica.objects.filter(receta=receta, fecha=fecha)
        if sucursal:
            existing_qs = existing_qs.filter(sucursal=sucursal)
        else:
            existing_qs = existing_qs.filter(sucursal__isnull=True)
        existing = existing_qs.order_by("id").first()

        if existing:
            if modo == "accumulate":
                existing.cantidad = Decimal(str(existing.cantidad or 0)) + cantidad
                existing.tickets = int(existing.tickets or 0) + tickets
                if monto_total > 0:
                    existing.monto_total = Decimal(str(existing.monto_total or 0)) + monto_total
            else:
                existing.cantidad = cantidad
                existing.tickets = tickets
                existing.monto_total = monto_total if monto_total > 0 else None
            existing.fuente = fuente
            existing.save(update_fields=["cantidad", "tickets", "monto_total", "fuente", "actualizado_en"])
            updated += 1
        else:
            VentaHistorica.objects.create(
                receta=receta,
                sucursal=sucursal,
                fecha=fecha,
                cantidad=cantidad,
                tickets=tickets,
                monto_total=monto_total if monto_total > 0 else None,
                fuente=fuente,
            )
            created += 1

    messages.success(
        request,
        f"Historial de ventas importado. Creados: {created}. Actualizados: {updated}. Omitidos: {skipped}.",
    )
    if unresolved_recetas:
        sample = ", ".join(unresolved_recetas[:5])
        extra = "" if len(unresolved_recetas) <= 5 else f" (+{len(unresolved_recetas) - 5} más)"
        messages.warning(request, f"Sin receta equivalente para: {sample}{extra}.")
    if unresolved_sucursales:
        sample = ", ".join(unresolved_sucursales[:5])
        extra = "" if len(unresolved_sucursales) <= 5 else f" (+{len(unresolved_sucursales) - 5} más)"
        messages.warning(request, f"Sucursales no encontradas: {sample}{extra}.")
    return redirect(next_url)


@login_required
@permission_required("recetas.add_planproduccion", raise_exception=True)
@require_POST
def pronostico_estadistico_desde_historial(request: HttpRequest) -> HttpResponse:
    plan_id = (request.POST.get("plan_id") or "").strip()
    next_params: dict[str, str] = {}
    if plan_id:
        next_params["plan_id"] = plan_id

    alcance = (request.POST.get("alcance") or "mes").strip().lower()
    if alcance not in {"mes", "semana", "fin_semana"}:
        alcance = "mes"
    periodo = _normalize_periodo_mes(request.POST.get("periodo"))
    fecha_base = _parse_date_safe(request.POST.get("fecha_base")) or timezone.localdate()
    incluir_preparaciones = request.POST.get("incluir_preparaciones") == "1"
    run_mode = (request.POST.get("run_mode") or "preview").strip().lower()
    if run_mode not in {"preview", "backtest", "apply_pronostico", "crear_plan", "aplicar_y_plan"}:
        run_mode = "preview"
    backtest_periods_raw = (request.POST.get("backtest_periods") or "3").strip()
    backtest_top_raw = (request.POST.get("backtest_top") or "10").strip()
    try:
        backtest_periods = int(backtest_periods_raw)
    except Exception:
        backtest_periods = 3
    try:
        backtest_top = int(backtest_top_raw)
    except Exception:
        backtest_top = 10
    backtest_periods = min(12, max(1, backtest_periods))
    backtest_top = min(30, max(1, backtest_top))
    escenario = (request.POST.get("escenario") or "base").strip().lower()
    if escenario not in {"base", "bajo", "alto"}:
        escenario = "base"
    escenario_to_key = {
        "base": "forecast_qty",
        "bajo": "forecast_low",
        "alto": "forecast_high",
    }
    qty_key = escenario_to_key.get(escenario, "forecast_qty")
    safety_pct = _to_decimal_safe(request.POST.get("safety_pct"))
    if safety_pct < Decimal("-30"):
        safety_pct = Decimal("-30")
    if safety_pct > Decimal("100"):
        safety_pct = Decimal("100")
    min_confianza_pct = _to_decimal_safe(request.POST.get("min_confianza_pct"))
    if min_confianza_pct < Decimal("0"):
        min_confianza_pct = Decimal("0")
    if min_confianza_pct > Decimal("100"):
        min_confianza_pct = Decimal("100")

    sucursal = Sucursal.objects.filter(pk=request.POST.get("sucursal_id")).first()

    resultado = _build_forecast_from_history(
        alcance=alcance,
        periodo=periodo,
        fecha_base=fecha_base,
        sucursal=sucursal,
        incluir_preparaciones=incluir_preparaciones,
        safety_pct=safety_pct,
    )
    resultado, filtered_conf = _filter_forecast_result_by_confianza(resultado, min_confianza_pct)
    resultado["min_confianza_pct"] = min_confianza_pct
    resultado["escenario"] = escenario
    request.session["pronostico_estadistico_preview"] = _forecast_session_payload(resultado, top_rows=120)
    if filtered_conf > 0:
        messages.info(
            request,
            f"Filtro de confianza >= {min_confianza_pct}% aplicado: {filtered_conf} recetas omitidas.",
        )
    if run_mode == "backtest":
        backtest_payload = _build_forecast_backtest_preview(
            alcance=alcance,
            fecha_base=fecha_base,
            periods=backtest_periods,
            sucursal=sucursal,
            incluir_preparaciones=incluir_preparaciones,
            safety_pct=safety_pct,
            min_confianza_pct=min_confianza_pct,
            escenario=escenario,
            top=backtest_top,
        )
        if backtest_payload is None:
            messages.warning(request, "No hay historial suficiente para backtest con esos filtros.")
            request.session["pronostico_backtest_preview"] = None
        else:
            request.session["pronostico_backtest_preview"] = backtest_payload
            messages.success(
                request,
                (
                    f"Backtest generado ({backtest_payload['totals']['windows_evaluated']} ventanas). "
                    f"MAPE promedio: {backtest_payload['totals']['mape_promedio'] if backtest_payload['totals']['mape_promedio'] is not None else '-'}%"
                ),
            )
        next_params["periodo"] = resultado["periodo"]
        next_params["forecast_compare_escenario"] = escenario
        if next_params:
            return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")
        return redirect(reverse("recetas:plan_produccion"))

    rows = resultado["rows"]
    if not rows:
        messages.warning(request, "No hay suficiente historial para generar pronóstico estadístico con esos filtros.")
        if next_params:
            return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")
        return redirect(reverse("recetas:plan_produccion"))

    created_forecast = 0
    updated_forecast = 0
    if run_mode in {"apply_pronostico", "aplicar_y_plan"}:
        if alcance != "mes":
            messages.warning(request, "Aplicar pronóstico mensual solo está disponible cuando el alcance es 'Mes'.")
        else:
            for row in rows:
                receta = Receta.objects.filter(pk=row["receta_id"]).first()
                if receta is None:
                    continue
                forecast_qty = Decimal(str(row.get(qty_key) or 0))
                current = PronosticoVenta.objects.filter(receta=receta, periodo=resultado["periodo"]).first()
                if current:
                    current.cantidad = forecast_qty
                    current.fuente = f"AUTO_HIST_{escenario.upper()}"[:40]
                    current.save(update_fields=["cantidad", "fuente", "actualizado_en"])
                    updated_forecast += 1
                else:
                    PronosticoVenta.objects.create(
                        receta=receta,
                        periodo=resultado["periodo"],
                        cantidad=forecast_qty,
                        fuente=f"AUTO_HIST_{escenario.upper()}"[:40],
                    )
                    created_forecast += 1

    created_plan = None
    if run_mode in {"crear_plan", "aplicar_y_plan"}:
        target_start = resultado["target_start"]
        default_name = f"Plan estadístico {resultado['target_start']} a {resultado['target_end']}"
        name = (request.POST.get("nombre_plan") or "").strip() or default_name
        plan = PlanProduccion.objects.create(
            nombre=name[:140],
            fecha_produccion=target_start,
            notas=(
                f"Generado por pronóstico estadístico ({alcance}) "
                f"{resultado['target_start']}..{resultado['target_end']} "
                f"- sucursal: {resultado['sucursal_nombre']}"
            )[:500],
            creado_por=request.user if request.user.is_authenticated else None,
        )
        lines = 0
        for row in rows:
            qty = Decimal(str(row.get(qty_key) or 0))
            if qty <= 0:
                continue
            PlanProduccionItem.objects.create(
                plan=plan,
                receta_id=row["receta_id"],
                cantidad=qty,
                notas="Pronóstico estadístico",
            )
            lines += 1
        if lines == 0:
            plan.delete()
            messages.warning(request, "No se creó plan: el pronóstico estadístico resultó en cantidades 0.")
        else:
            created_plan = plan
            next_params["plan_id"] = str(plan.id)
            messages.success(request, f"Plan estadístico creado con {lines} renglones.")

    if run_mode == "preview":
        messages.success(request, f"Vista previa generada ({len(rows)} recetas) para {resultado['sucursal_nombre']}.")
    if created_forecast or updated_forecast:
        messages.success(
            request,
            (
                f"Pronóstico mensual aplicado desde historial ({escenario}). "
                f"Creados: {created_forecast}. Actualizados: {updated_forecast}."
            ),
        )
    if run_mode in {"crear_plan", "aplicar_y_plan"} and created_plan is None:
        messages.info(request, "Se generó la vista previa estadística, pero no se creó plan.")

    next_params["periodo"] = resultado["periodo"]
    next_params["forecast_compare_escenario"] = escenario
    if next_params:
        return redirect(f"{reverse('recetas:plan_produccion')}?{urlencode(next_params)}")
    return redirect(reverse("recetas:plan_produccion"))


@login_required
@permission_required("recetas.change_planproduccion", raise_exception=True)
@require_POST
def pronosticos_importar(request: HttpRequest) -> HttpResponse:
    uploaded = request.FILES.get("archivo")
    plan_id = (request.POST.get("plan_id") or "").strip()
    modo = (request.POST.get("modo") or "replace").strip().lower()
    if modo not in {"replace", "accumulate"}:
        modo = "replace"
    fuente = (request.POST.get("fuente") or "UI_PRONOSTICO").strip()[:40] or "UI_PRONOSTICO"
    periodo_default = _normalize_periodo_mes(request.POST.get("periodo_default"))

    next_url = reverse("recetas:plan_produccion")
    if plan_id:
        next_url = f"{next_url}?{urlencode({'plan_id': plan_id})}"

    if not uploaded:
        messages.error(request, "Selecciona un archivo para importar pronósticos.")
        return redirect(next_url)

    try:
        rows = _load_pronostico_rows(uploaded)
    except Exception as exc:
        messages.error(request, f"No se pudo leer el archivo: {exc}")
        return redirect(next_url)

    if not rows:
        messages.warning(request, "Archivo sin filas válidas para pronósticos.")
        return redirect(next_url)

    created = 0
    updated = 0
    skipped = 0
    unresolved: list[str] = []

    for row in rows:
        receta_name = (row.get("receta") or "").strip()
        codigo_point = (row.get("codigo_point") or "").strip()
        cantidad = _to_decimal_safe(row.get("cantidad"))
        periodo = _normalize_periodo_mes(str(row.get("periodo") or row.get("mes") or periodo_default))

        if cantidad < 0:
            skipped += 1
            continue
        if not receta_name and not codigo_point:
            skipped += 1
            continue

        receta = None
        if codigo_point:
            receta = Receta.objects.filter(codigo_point__iexact=codigo_point).order_by("id").first()
        if receta is None and receta_name:
            receta = Receta.objects.filter(nombre_normalizado=normalizar_nombre(receta_name)).order_by("id").first()

        if receta is None:
            unresolved.append(receta_name or codigo_point)
            skipped += 1
            continue

        pronostico = PronosticoVenta.objects.filter(receta=receta, periodo=periodo).first()
        if pronostico:
            if modo == "accumulate":
                pronostico.cantidad = Decimal(str(pronostico.cantidad or 0)) + cantidad
            else:
                pronostico.cantidad = cantidad
            pronostico.fuente = fuente
            pronostico.save(update_fields=["cantidad", "fuente", "actualizado_en"])
            updated += 1
        else:
            PronosticoVenta.objects.create(
                receta=receta,
                periodo=periodo,
                cantidad=cantidad,
                fuente=fuente,
            )
            created += 1

    messages.success(
        request,
        f"Pronósticos importados. Creados: {created}. Actualizados: {updated}. Omitidos: {skipped}. Modo: {'acumular' if modo == 'accumulate' else 'reemplazar'}.",
    )
    if unresolved:
        sample = ", ".join(unresolved[:5])
        extra = "" if len(unresolved) <= 5 else f" (+{len(unresolved) - 5} más)"
        messages.warning(
            request,
            f"Sin receta equivalente para: {sample}{extra}. Revisa nombre/código comercial y vuelve a importar.",
        )
    return redirect(next_url)


@login_required
@permission_required("recetas.add_planproduccion", raise_exception=True)
@require_POST
def plan_produccion_generar_desde_pronostico(request: HttpRequest) -> HttpResponse:
    periodo = _normalize_periodo_mes(request.POST.get("periodo"))
    nombre = (request.POST.get("nombre") or "").strip()
    fecha_raw = (request.POST.get("fecha_produccion") or "").strip()
    incluir_preparaciones = request.POST.get("incluir_preparaciones") == "1"

    if not nombre:
        nombre = f"Plan desde pronóstico {periodo}"

    try:
        fecha_plan = date.fromisoformat(fecha_raw) if fecha_raw else date.fromisoformat(f"{periodo}-01")
    except Exception:
        messages.error(request, "Fecha de producción inválida.")
        return redirect(f"{reverse('recetas:plan_produccion')}?periodo={periodo}")

    pronosticos_qs = PronosticoVenta.objects.filter(periodo=periodo).select_related("receta").order_by("receta__nombre")
    if not incluir_preparaciones:
        pronosticos_qs = pronosticos_qs.filter(receta__tipo=Receta.TIPO_PRODUCTO_FINAL)

    pronosticos = list(pronosticos_qs)
    if not pronosticos:
        messages.warning(
            request,
            "No hay pronósticos para generar plan en ese período con los filtros actuales.",
        )
        return redirect(f"{reverse('recetas:plan_produccion')}?periodo={periodo}")

    plan = PlanProduccion.objects.create(
        nombre=nombre[:140],
        fecha_produccion=fecha_plan,
        notas=f"Generado desde pronóstico {periodo}",
        creado_por=request.user if request.user.is_authenticated else None,
    )

    created = 0
    skipped = 0
    for p in pronosticos:
        qty = Decimal(str(p.cantidad or 0))
        if qty <= 0:
            skipped += 1
            continue
        PlanProduccionItem.objects.create(
            plan=plan,
            receta=p.receta,
            cantidad=qty,
            notas=f"Pronóstico {periodo}",
        )
        created += 1

    if created == 0:
        plan.delete()
        messages.warning(
            request,
            "No se creó plan: todos los pronósticos tenían cantidad 0.",
        )
        return redirect(f"{reverse('recetas:plan_produccion')}?periodo={periodo}")

    messages.success(
        request,
        f"Plan generado desde pronóstico {periodo}. Renglones creados: {created}. Omitidos por cantidad 0: {skipped}.",
    )
    return redirect(f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}&periodo={periodo}")


def _plan_trunk_handoff_rows(
    *,
    plan_actual: PlanProduccion | None,
    explosion: Dict[str, Any] | None,
    document_control: dict[str, Any] | None,
    demand_gate_summary: dict[str, Any] | None = None,
    master_demand_gate_summary: dict[str, Any] | None = None,
) -> list[dict[str, object]]:
    master_blockers = int((document_control or {}).get("master_blocker_total") or 0)
    document_blockers = int((document_control or {}).get("blocked_total") or 0)
    lineas_abiertas = int((explosion or {}).get("lineas_sin_match") or 0) + len((explosion or {}).get("lineas_sin_cantidad") or [])
    costo_abierto = len((explosion or {}).get("lineas_sin_costo_unitario") or [])
    capacidad_abierta = int((explosion or {}).get("alertas_capacidad") or 0)
    demand_blockers = int((demand_gate_summary or {}).get("blockers") or 0)
    master_demand_blockers = int((master_demand_gate_summary or {}).get("blockers") or 0)
    plan_blockers = master_blockers + lineas_abiertas + costo_abierto + demand_blockers + master_demand_blockers
    compras_blockers = master_blockers + document_blockers + master_demand_blockers
    inventario_blockers = master_blockers + capacidad_abierta + master_demand_blockers
    critical_master_open = master_demand_blockers > 0
    plan_url = (
        f"{reverse('recetas:plan_produccion')}?{urlencode({'plan_id': plan_actual.id})}"
        if plan_actual
        else reverse("recetas:plan_produccion")
    )

    compras_ready = bool(
        plan_actual
        and compras_blockers == 0
        and int((document_control or {}).get("solicitudes_total") or 0) > 0
    )
    inventario_ready = bool(plan_actual and inventario_blockers == 0)
    plan_ready = bool(plan_actual and plan_blockers == 0)

    return [
        {
            "label": "Plan / BOM",
            "owner": "Producción / Costeo",
            "status": "Crítico" if critical_master_open else "Listo para operar" if plan_ready else "Bloqueado",
            "tone": "danger" if critical_master_open else "success" if plan_ready else "warning",
            "blockers": plan_blockers if plan_actual else 1,
            "completion": 12 if critical_master_open else 100 if plan_ready else max(0, 100 - ((plan_blockers or 1) * 8)),
            "depends_on": "Plan activo + estructura BOM + costo estable + señal comercial",
            "exit_criteria": "El plan debe explotar materiales y costos sin partidas abiertas, con artículos maestros completos y base comercial utilizable.",
            "detail": (
                "El plan no debe liberarse mientras siga abierta la demanda crítica bloqueada por maestro."
                if critical_master_open
                else "El plan ya puede operar como documento rector de producción."
                if plan_ready
                else "Todavía faltan cierres en plan, BOM, costo, catálogo ERP o prioridades críticas del maestro."
            ),
            "next_step": (
                (master_demand_gate_summary or {}).get("next_step")
                if master_demand_blockers
                else (demand_gate_summary or {}).get("next_step")
                if demand_blockers
                else ("Abrir plan" if plan_actual else "Crear plan")
            ),
            "url": (
                (master_demand_gate_summary or {}).get("action_url")
                if master_demand_blockers
                else (demand_gate_summary or {}).get("action_url")
                if demand_blockers
                else plan_url
            ),
            "cta": (
                (master_demand_gate_summary or {}).get("action_label")
                if master_demand_blockers
                else (demand_gate_summary or {}).get("action_label")
                if demand_blockers
                else ("Abrir plan" if plan_actual else "Crear plan")
            ),
        },
        {
            "label": "Compras documentales",
            "owner": "Compras / Planeación",
            "status": "Crítico" if critical_master_open else "Listo para operar" if compras_ready else "Bloqueado",
            "tone": "danger" if critical_master_open else "success" if compras_ready else "warning",
            "blockers": compras_blockers if plan_actual else 1,
            "completion": 18 if critical_master_open else 100 if compras_ready else max(0, 100 - ((compras_blockers or 1) * 10)),
            "depends_on": "Solicitudes + órdenes + recepciones sin bloqueo",
            "exit_criteria": "El flujo documental debe quedar emitido y sin bloqueos antes de abastecer.",
            "detail": (
                "Compras no debe emitir documentos mientras siga abierta la demanda crítica bloqueada por maestro."
                if critical_master_open
                else "Compras ya puede operar el plan como documento de abastecimiento."
                if compras_ready
                else "Compras sigue condicionado por bloqueos documentales o por la falta de documentos."
            ),
            "next_step": (
                (master_demand_gate_summary or {}).get("next_step")
                if critical_master_open
                else (document_control or {}).get("next_action_label") or "Abrir compras"
            ),
            "url": (
                (master_demand_gate_summary or {}).get("action_url")
                if critical_master_open
                else (document_control or {}).get("next_action_url") or reverse("compras:solicitudes")
            ),
            "cta": (
                (master_demand_gate_summary or {}).get("action_label")
                if critical_master_open
                else (document_control or {}).get("next_action_label") or "Abrir compras"
            ),
        },
        {
            "label": "Inventario / Reabasto",
            "owner": "Inventario / Almacén",
            "status": "Listo para operar" if inventario_ready else "Bloqueado",
            "tone": "success" if inventario_ready else "warning",
            "blockers": inventario_blockers if plan_actual else 1,
            "completion": 100 if inventario_ready else max(0, 100 - ((inventario_blockers or 1) * 12)),
            "depends_on": "Stock operativo + artículos maestros completos",
            "exit_criteria": "Inventario debe sostener faltantes, stock y reabasto del plan sin brechas de maestro.",
            "detail": (
                "Inventario ya puede sostener el plan sin alertas de capacidad ni brechas de maestro."
                if inventario_ready
                else "Inventario todavía requiere cierre de catálogo ERP o faltantes de stock para operar estable."
            ),
            "next_step": "Abrir inventario" if plan_actual else "Esperar plan",
            "url": reverse("inventario:existencias"),
            "cta": "Abrir inventario",
        },
    ]


def _ventas_historicas_plan_summary() -> dict[str, Any]:
    rows_qs = VentaHistorica.objects.select_related("sucursal", "receta")
    total_rows = rows_qs.count()
    if total_rows == 0:
        return {
            "available": False,
            "status": "Sin histórico",
            "tone": "warning",
            "detail": "Todavía no hay base diaria cargada para apoyar el plan.",
            "date_label": "Sin cobertura",
            "active_days": 0,
            "expected_days": 0,
            "missing_days": 0,
            "branch_count": 0,
            "recipe_count": 0,
            "total_rows": 0,
            "total_units": Decimal("0"),
            "top_branches": [],
            "top_recipes": [],
        }

    first_date = rows_qs.order_by("fecha").values_list("fecha", flat=True).first()
    last_date = rows_qs.order_by("-fecha").values_list("fecha", flat=True).first()
    active_days = rows_qs.values_list("fecha", flat=True).distinct().count()
    expected_days = ((last_date - first_date).days + 1) if first_date and last_date else 0
    missing_days = max(expected_days - active_days, 0)
    total_units = rows_qs.aggregate(total=Sum("cantidad")).get("total") or Decimal("0")
    branch_count = rows_qs.exclude(sucursal_id__isnull=True).values_list("sucursal_id", flat=True).distinct().count()
    recipe_count = rows_qs.values_list("receta_id", flat=True).distinct().count()
    top_branches = list(
        rows_qs.exclude(sucursal_id__isnull=True)
        .values("sucursal__codigo", "sucursal__nombre")
        .annotate(total=Sum("cantidad"))
        .order_by("-total", "sucursal__codigo")[:4]
    )
    top_recipes = list(
        rows_qs.values("receta__nombre")
        .annotate(total=Sum("cantidad"))
        .order_by("-total", "receta__nombre")[:5]
    )
    return {
        "available": True,
        "status": "Cobertura cerrada" if missing_days == 0 else "Cobertura parcial",
        "tone": "success" if missing_days == 0 else "warning",
        "detail": (
            "La base diaria está lista para pronóstico y comparación contra plan."
            if missing_days == 0
            else f"Hay {missing_days} día(s) faltantes dentro del rango histórico cargado."
        ),
        "date_label": (
            f"{first_date.strftime('%d/%m/%Y')} → {last_date.strftime('%d/%m/%Y')}"
            if first_date and last_date
            else "Sin cobertura"
        ),
        "active_days": active_days,
        "expected_days": expected_days,
        "missing_days": missing_days,
        "branch_count": branch_count,
        "recipe_count": recipe_count,
        "total_rows": total_rows,
        "total_units": total_units,
        "top_branches": top_branches,
        "top_recipes": top_recipes,
    }


def _reabasto_demand_history_summary(fecha_operacion: date) -> dict[str, Any]:
    weekday = fecha_operacion.weekday()
    base_qs = VentaHistorica.objects.select_related("sucursal", "receta").filter(
        fecha__lt=fecha_operacion,
        fecha__week_day=((weekday + 2) if weekday < 6 else 1),
    )
    rows_qs = base_qs.order_by("-fecha")[:4000]
    if not rows_qs:
        rows_qs = VentaHistorica.objects.select_related("sucursal", "receta").filter(fecha__lt=fecha_operacion).order_by("-fecha")[:4000]
    if not rows_qs:
        return {
            "available": False,
            "status": "Sin base comparable",
            "tone": "warning",
            "detail": "Todavía no hay ventas históricas suficientes para orientar el reabasto.",
            "scope_label": "Sin cobertura",
            "sample_days": 0,
            "branch_count": 0,
            "recipe_count": 0,
            "total_units": Decimal("0"),
            "top_branches": [],
            "top_recipes": [],
            "years_observed": 0,
            "comparable_years": 0,
        }

    row_ids = [row.id for row in rows_qs]
    rows_agg = VentaHistorica.objects.filter(id__in=row_ids)
    history_meta = _forecast_history_meta(
        base_qs if base_qs.exists() else VentaHistorica.objects.filter(fecha__lt=fecha_operacion),
        alcance="semana",
        target_start=fecha_operacion,
        target_end=fecha_operacion,
    )
    first_date = rows_agg.order_by("fecha").values_list("fecha", flat=True).first()
    last_date = rows_agg.order_by("-fecha").values_list("fecha", flat=True).first()
    sample_days = rows_agg.values_list("fecha", flat=True).distinct().count()
    branch_count = rows_agg.exclude(sucursal_id__isnull=True).values_list("sucursal_id", flat=True).distinct().count()
    recipe_count = rows_agg.values_list("receta_id", flat=True).distinct().count()
    total_units = rows_agg.aggregate(total=Sum("cantidad")).get("total") or Decimal("0")
    top_branches = list(
        rows_agg.exclude(sucursal_id__isnull=True)
        .values("sucursal__codigo", "sucursal__nombre")
        .annotate(total=Sum("cantidad"))
        .order_by("-total", "sucursal__codigo")[:4]
    )
    top_recipes = list(
        rows_agg.values("receta__nombre")
        .annotate(total=Sum("cantidad"))
        .order_by("-total", "receta__nombre")[:5]
    )
    return {
        "available": True,
        "status": "Base comparable multianual" if int(history_meta.get("comparable_years") or 0) >= 3 else "Base comparable lista",
        "tone": "success",
        "detail": (
            "Usa temporadas comparables de varios años como referencia para cierres de sucursal y faltante CEDIS."
            if int(history_meta.get("comparable_years") or 0) >= 3
            else "Usa este histórico comparable como referencia para cierres de sucursal y faltante CEDIS."
        ),
        "scope_label": (
            f"{first_date.strftime('%d/%m/%Y')} → {last_date.strftime('%d/%m/%Y')}"
            if first_date and last_date
            else "Sin cobertura"
        ),
        "sample_days": sample_days,
        "branch_count": branch_count,
        "recipe_count": recipe_count,
        "total_units": total_units,
        "top_branches": top_branches,
        "top_recipes": top_recipes,
        "years_observed": int(history_meta.get("years_observed") or 0),
        "comparable_years": int(history_meta.get("comparable_years") or 0),
    }


@login_required
@permission_required("recetas.view_planproduccion", raise_exception=True)
def plan_produccion(request: HttpRequest) -> HttpResponse:
    estado_plan = (request.GET.get("estado_plan") or "all").strip().lower()
    estado_plan_map = {
        "all": None,
        "borrador": PlanProduccion.ESTADO_BORRADOR,
        "consumo_aplicado": PlanProduccion.ESTADO_CONSUMO_APLICADO,
        "cerrado": PlanProduccion.ESTADO_CERRADO,
    }
    if estado_plan not in estado_plan_map:
        estado_plan = "all"
    dg_filters = _plan_status_dashboard_filters(request)

    planes_qs = PlanProduccion.objects.select_related("creado_por").prefetch_related("items").order_by("-fecha_produccion", "-id")
    plan_status_dashboard = _plan_status_dashboard(
        planes_qs,
        start_date=dg_filters["start_date"],
        end_date=dg_filters["end_date"],
        group_by=dg_filters["group_by"],
        limit=12,
    )
    plan_status_cards = [
        {
            "key": "all",
            "label": "Todos",
            "count": planes_qs.count(),
            "tone": "primary",
        },
        {
            "key": "borrador",
            "label": "Borrador",
            "count": planes_qs.filter(estado=PlanProduccion.ESTADO_BORRADOR).count(),
            "tone": "warning",
        },
        {
            "key": "consumo_aplicado",
            "label": "Consumo aplicado",
            "count": planes_qs.filter(estado=PlanProduccion.ESTADO_CONSUMO_APLICADO).count(),
            "tone": "primary",
        },
        {
            "key": "cerrado",
            "label": "Cerrado",
            "count": planes_qs.filter(estado=PlanProduccion.ESTADO_CERRADO).count(),
            "tone": "success",
        },
    ]
    planes = planes_qs
    estado_db = estado_plan_map.get(estado_plan)
    if estado_db:
        planes = planes.filter(estado=estado_db)
    plan_actual = None
    plan_id = request.GET.get("plan_id")
    if plan_id:
        plan_actual = get_object_or_404(PlanProduccion, pk=plan_id)
    elif planes.exists():
        plan_actual = planes.first()

    recetas_disponibles = Receta.objects.order_by("tipo", "nombre")
    explosion = _plan_explosion(plan_actual) if plan_actual else None
    plan_vs_pronostico = _plan_vs_pronostico(plan_actual) if plan_actual else None
    periodo_pronostico_default = _normalize_periodo_mes(request.GET.get("periodo"))
    mrp_periodo = _normalize_periodo_mes(request.GET.get("mrp_periodo"))
    mrp_periodo_tipo = (request.GET.get("mrp_periodo_tipo") or "mes").strip().lower()
    alcance_estadistico = (request.GET.get("alcance_estadistico") or "mes").strip().lower()
    if alcance_estadistico not in {"mes", "semana", "fin_semana"}:
        alcance_estadistico = "mes"
    fecha_base_estadistica = request.GET.get("fecha_base_estadistica") or timezone.localdate().isoformat()
    backtest_periods = request.GET.get("backtest_periods") or "3"
    backtest_top = request.GET.get("backtest_top") or "10"
    if mrp_periodo_tipo not in {"mes", "q1", "q2"}:
        mrp_periodo_tipo = "mes"
    mrp_focus_kind = (request.GET.get("mrp_focus_kind") or "").strip().lower()
    mrp_focus_key = (request.GET.get("mrp_focus_key") or "").strip().lower()
    mrp_periodo_resumen = _periodo_mrp_resumen(
        mrp_periodo,
        mrp_periodo_tipo,
        focus_kind=mrp_focus_kind,
        focus_key=mrp_focus_key,
    )
    pronosticos_unavailable = False
    ventas_historicas_unavailable = False
    solicitudes_venta_unavailable = False
    try:
        pronosticos_periodo = PronosticoVenta.objects.filter(periodo=periodo_pronostico_default)
        pronosticos_periodo_count = pronosticos_periodo.count()
        pronosticos_periodo_total = pronosticos_periodo.aggregate(total=Sum("cantidad")).get("total") or Decimal("0")
    except (OperationalError, ProgrammingError):
        pronosticos_periodo_count = 0
        pronosticos_periodo_total = Decimal("0")
        pronosticos_unavailable = True
    try:
        ventas_historicas_count = VentaHistorica.objects.count()
        ventas_historicas_total = VentaHistorica.objects.aggregate(total=Sum("cantidad")).get("total") or Decimal("0")
        ventas_hist_fecha_max = VentaHistorica.objects.order_by("-fecha").values_list("fecha", flat=True).first()
        ventas_historicas_summary = _ventas_historicas_plan_summary()
    except (OperationalError, ProgrammingError):
        ventas_historicas_count = 0
        ventas_historicas_total = Decimal("0")
        ventas_hist_fecha_max = None
        ventas_historicas_summary = {
            "available": False,
            "status": "Sin histórico",
            "tone": "warning",
            "detail": "La base diaria no está disponible en este entorno.",
            "date_label": "Sin cobertura",
            "active_days": 0,
            "expected_days": 0,
            "missing_days": 0,
            "branch_count": 0,
            "recipe_count": 0,
            "total_rows": 0,
            "total_units": Decimal("0"),
            "top_branches": [],
            "top_recipes": [],
        }
        ventas_historicas_unavailable = True
    try:
        solicitudes_venta_periodo = SolicitudVenta.objects.filter(periodo=periodo_pronostico_default)
        solicitudes_venta_count = solicitudes_venta_periodo.count()
        solicitudes_venta_total = solicitudes_venta_periodo.aggregate(total=Sum("cantidad")).get("total") or Decimal("0")
        solicitudes_venta_fecha_max = solicitudes_venta_periodo.order_by("-fecha_inicio").values_list("fecha_inicio", flat=True).first()
    except (OperationalError, ProgrammingError):
        solicitudes_venta_count = 0
        solicitudes_venta_total = Decimal("0")
        solicitudes_venta_fecha_max = None
        solicitudes_venta_unavailable = True
    forecast_preview = request.session.get("pronostico_estadistico_preview")
    forecast_backtest = request.session.get("pronostico_backtest_preview")
    min_confianza_default = request.GET.get("min_confianza_pct")
    if not min_confianza_default:
        min_confianza_default = str((forecast_preview or {}).get("min_confianza_pct") or "0")
    forecast_run_escenario_default = str((forecast_preview or {}).get("escenario") or "base").strip().lower()
    if forecast_run_escenario_default not in {"base", "bajo", "alto"}:
        forecast_run_escenario_default = "base"
    forecast_compare_escenario = (request.GET.get("forecast_compare_escenario") or "").strip().lower()
    if forecast_compare_escenario not in {"base", "bajo", "alto"}:
        forecast_compare_escenario = str((forecast_preview or {}).get("escenario") or "base").strip().lower()
    if forecast_compare_escenario not in {"base", "bajo", "alto"}:
        forecast_compare_escenario = "base"
    try:
        forecast_vs_solicitud = _forecast_vs_solicitud_preview(
            forecast_preview,
            escenario=forecast_compare_escenario,
        )
    except (OperationalError, ProgrammingError):
        forecast_vs_solicitud = None
        solicitudes_venta_unavailable = True
    forecast_preview_summary = _forecast_preview_operational_summary(forecast_preview)
    demand_gate_summary = _commercial_signal_gate(
        forecast_preview_summary,
        context_label="el plan de producción",
        action_url="#plan-pronosticos",
        action_label="Abrir pronóstico",
    )
    master_demand_gate_summary = _plan_master_demand_gate(plan_actual)
    sucursales = sucursales_operativas()
    enterprise_board = _plan_enterprise_board(
        plan_actual,
        explosion,
        plan_vs_pronostico,
        mrp_periodo_resumen,
    )
    stage_key = (request.GET.get("stage_key") or "auto").strip().lower()
    closure_key = (request.GET.get("closure_key") or "auto").strip().lower()
    handoff_key = (request.GET.get("handoff_key") or "auto").strip().lower()
    master_focus_key = (request.GET.get("master_focus_key") or "auto").strip().lower()
    master_missing_key = (request.GET.get("master_missing_key") or "auto").strip().lower()
    document_control = _plan_document_control(
        plan_actual,
        stage_key=stage_key,
        closure_key=closure_key,
        handoff_key=handoff_key,
        master_focus_key=master_focus_key,
        master_missing_key=master_missing_key,
    )
    if document_control and enterprise_board:
        plan_focus_base = reverse("recetas:plan_produccion") + f"?{urlencode({'plan_id': plan_actual.id})}"
        master_cards = list(document_control.get("master_blocker_class_cards", [])[:4])
        master_missing_cards = list(document_control.get("master_blocker_missing_cards", [])[:4])
        master_detail_rows = list(document_control.get("master_blocker_detail_rows", [])[:8])
        valid_master_focus_keys = {
            str(card.get("key") or "").strip().lower() for card in master_cards if card.get("key")
        }
        valid_master_missing_keys = {
            str(card.get("key") or "").strip().lower() for card in master_missing_cards if card.get("key")
        }
        selected_master_focus_key = document_control.get("selected_master_focus_key") or "auto"
        selected_master_missing_key = document_control.get("selected_master_missing_key") or "auto"
        if selected_master_focus_key not in valid_master_focus_keys:
            selected_master_focus_key = "auto"
        if selected_master_missing_key not in valid_master_missing_keys:
            selected_master_missing_key = "auto"
        for card in master_cards:
            card_key = str(card.get("key") or "").strip().lower()
            card["focus_url"] = f"{plan_focus_base}&master_focus_key={urlencode({'k': card_key})[2:]}"
            card["is_active"] = selected_master_focus_key != "auto" and card_key == selected_master_focus_key
        for card in master_missing_cards:
            card_key = str(card.get("key") or "").strip().lower()
            card["focus_url"] = f"{plan_focus_base}&master_missing_key={urlencode({'k': card_key})[2:]}"
            card["is_active"] = selected_master_missing_key != "auto" and card_key == selected_master_missing_key
        filtered_master_rows = (
            [row for row in master_detail_rows if str(row.get("class_key") or "").strip().lower() == selected_master_focus_key]
            if selected_master_focus_key != "auto"
            else master_detail_rows
        )
        if selected_master_missing_key != "auto":
            filtered_master_rows = [
                row
                for row in filtered_master_rows
                if str(row.get("missing_field") or "").strip().lower() == selected_master_missing_key
            ]
        document_control["selected_master_focus_key"] = selected_master_focus_key
        document_control["selected_master_missing_key"] = selected_master_missing_key
        document_control["master_blocker_class_cards"] = master_cards
        document_control["master_blocker_missing_cards"] = master_missing_cards
        document_control["master_blocker_detail_rows"] = filtered_master_rows
        master_focus_rows: list[dict[str, Any]] = []
        for row in document_control["master_blocker_detail_rows"][:3]:
            master_focus_rows.append(
                {
                    "class_key": row.get("class_key", ""),
                    "class_label": row.get("class_label", ""),
                    "missing_key": row.get("missing_field", ""),
                    "insumo_nombre": row.get("name", ""),
                    "missing_field": row.get("missing", ""),
                    "detail": row.get("detail", ""),
                    "action_label": row.get("action_label", "Abrir maestro"),
                    "action_url": row.get("action_url", reverse("maestros:insumo_list")),
                    "edit_url": row.get("edit_url", ""),
                    "action_detail": row.get("action_detail", ""),
                    "tone": "warning",
                }
            )
        document_control["master_focus_rows"] = master_focus_rows
        if master_focus_rows:
            first_master_focus = master_focus_rows[0]
            document_control["master_focus"] = {
                **first_master_focus,
                "label": f"{first_master_focus['class_label']} · {first_master_focus['missing_field']}",
                "summary": (
                    f"El flujo del plan sigue bloqueado por {first_master_focus['insumo_nombre']} "
                    f"({first_master_focus['missing_field']})."
                ),
                "tone": "warning",
            }
        else:
            document_control["master_focus"] = None
    critical_path_rows = _recipes_critical_path_rows(
        document_control["document_stage_rows"] if document_control else [],
        owner="Plan / Compras / Producción",
        fallback_url=reverse("recetas:plan_produccion"),
    )
    trunk_handoff_rows = _plan_trunk_handoff_rows(
        plan_actual=plan_actual,
        explosion=explosion,
        document_control=document_control,
        demand_gate_summary=demand_gate_summary,
        master_demand_gate_summary=master_demand_gate_summary,
    )
    critical_master_demand_rows = list((master_demand_gate_summary or {}).get("rows") or [])[:3]
    daily_decision_rows = _plan_daily_decisions(
        plan_actual=plan_actual,
        demand_gate_summary=demand_gate_summary,
        master_demand_gate_summary=master_demand_gate_summary,
        ventas_historicas_summary=ventas_historicas_summary,
        document_control=document_control,
    )
    branch_priority_rows = _plan_branch_priority_rows(
        plan_actual=plan_actual,
        periodo=periodo_pronostico_default,
    )
    branch_supply_rows = _plan_branch_supply_rows(
        branch_priority_rows=branch_priority_rows,
    )
    return render(
        request,
        "recetas/plan_produccion.html",
        {
            "planes": planes[:30],
            "plan_status_dashboard": plan_status_dashboard,
            "plan_status_cards": plan_status_cards,
            "selected_plan_status_filter": estado_plan,
            "dg_start_date": dg_filters["start_date"].isoformat() if dg_filters["start_date"] else "",
            "dg_end_date": dg_filters["end_date"].isoformat() if dg_filters["end_date"] else "",
            "dg_group_by": dg_filters["group_by"],
            "plan_actual": plan_actual,
            "recetas_disponibles": recetas_disponibles,
            "explosion": explosion,
            "plan_vs_pronostico": plan_vs_pronostico,
            "periodo_pronostico_default": periodo_pronostico_default,
            "mrp_periodo": mrp_periodo,
            "mrp_periodo_tipo": mrp_periodo_tipo,
            "mrp_focus_kind": mrp_focus_kind,
            "mrp_focus_key": mrp_focus_key,
            "mrp_periodo_resumen": mrp_periodo_resumen,
            "pronosticos_periodo_count": pronosticos_periodo_count,
            "pronosticos_periodo_total": pronosticos_periodo_total,
            "pronosticos_unavailable": pronosticos_unavailable,
            "ventas_historicas_count": ventas_historicas_count,
            "ventas_historicas_total": ventas_historicas_total,
            "ventas_hist_fecha_max": ventas_hist_fecha_max,
            "ventas_historicas_summary": ventas_historicas_summary,
            "ventas_historicas_unavailable": ventas_historicas_unavailable,
            "solicitudes_venta_count": solicitudes_venta_count,
            "solicitudes_venta_total": solicitudes_venta_total,
            "solicitudes_venta_fecha_max": solicitudes_venta_fecha_max,
            "solicitudes_venta_unavailable": solicitudes_venta_unavailable,
            "forecast_preview": forecast_preview,
            "forecast_backtest": forecast_backtest,
            "forecast_vs_solicitud": forecast_vs_solicitud,
            "forecast_preview_summary": forecast_preview_summary,
            "forecast_vs_solicitud_summary": _forecast_vs_solicitud_operational_summary(forecast_vs_solicitud),
            "demand_gate_summary": demand_gate_summary,
            "master_demand_gate_summary": master_demand_gate_summary,
            "critical_master_demand_rows": critical_master_demand_rows,
            "daily_decision_rows": daily_decision_rows,
            "branch_priority_rows": branch_priority_rows,
            "branch_supply_rows": branch_supply_rows,
            "forecast_compare_escenario": forecast_compare_escenario,
            "forecast_run_escenario_default": forecast_run_escenario_default,
            "sucursales": sucursales,
            "alcance_estadistico": alcance_estadistico,
            "fecha_base_estadistica": fecha_base_estadistica,
            "backtest_periods": backtest_periods,
            "backtest_top": backtest_top,
            "min_confianza_default": min_confianza_default,
            "enterprise_board": enterprise_board,
            "document_control": document_control,
            "executive_radar_rows": _recipes_executive_radar_rows(
                document_control["document_stage_rows"] if document_control else [],
                owner="Plan / Compras / Producción",
                fallback_url=reverse("recetas:plan_produccion"),
            ),
            "critical_path_rows": critical_path_rows,
            "trunk_handoff_rows": trunk_handoff_rows,
            "trunk_handoff_summary": _trunk_handoff_summary(
                trunk_handoff_rows,
                owner="Plan / Compras / Producción",
                fallback_url=reverse("recetas:plan_produccion"),
            ),
        },
    )


@login_required
@permission_required("recetas.view_planproduccion", raise_exception=True)
def plan_produccion_export(request: HttpRequest, plan_id: int) -> HttpResponse:
    plan = get_object_or_404(PlanProduccion, pk=plan_id)
    export_format = (request.GET.get("format") or "csv").lower()
    if export_format in {"point", "point-xlsx", "point_xlsx"}:
        return _export_plan_point_xlsx(plan)
    explosion = _plan_explosion(plan)
    if export_format == "xlsx":
        return _export_plan_xlsx(plan, explosion)
    return _export_plan_csv(plan, explosion)


@login_required
@permission_required("recetas.view_planproduccion", raise_exception=True)
def plan_produccion_periodo_export(request: HttpRequest) -> HttpResponse:
    periodo = _normalize_periodo_mes(request.GET.get("mrp_periodo"))
    periodo_tipo = (request.GET.get("mrp_periodo_tipo") or "mes").strip().lower()
    if periodo_tipo not in {"mes", "q1", "q2"}:
        periodo_tipo = "mes"
    export_format = (request.GET.get("format") or "csv").strip().lower()
    resumen = _periodo_mrp_resumen(periodo, periodo_tipo)
    if export_format == "xlsx":
        return _export_periodo_mrp_xlsx(resumen)
    return _export_periodo_mrp_csv(resumen)


@login_required
@permission_required("recetas.view_planproduccion", raise_exception=True)
def plan_produccion_estado_dashboard_export(request: HttpRequest) -> HttpResponse:
    export_format = (request.GET.get("format") or "csv").strip().lower()
    dg_filters = _plan_status_dashboard_filters(request)
    summary = _plan_status_dashboard(
        PlanProduccion.objects.all().order_by("-fecha_produccion", "-id"),
        start_date=dg_filters["start_date"],
        end_date=dg_filters["end_date"],
        group_by=dg_filters["group_by"],
        limit=52 if dg_filters["group_by"] == "week" else 36 if dg_filters["group_by"] == "month" else 31,
    )
    if export_format == "xlsx":
        return _export_plan_status_dashboard_xlsx(summary)
    return _export_plan_status_dashboard_csv(summary)


@login_required
@permission_required("recetas.view_planproduccion", raise_exception=True)
def plan_produccion_dg_dashboard(request: HttpRequest) -> HttpResponse:
    dg_filters = _plan_status_dashboard_filters(request)
    planes_qs = PlanProduccion.objects.select_related("creado_por").prefetch_related("items").order_by("-fecha_produccion", "-id")
    planes_corte_qs = planes_qs
    if dg_filters["start_date"]:
        planes_corte_qs = planes_corte_qs.filter(fecha_produccion__gte=dg_filters["start_date"])
    if dg_filters["end_date"]:
        planes_corte_qs = planes_corte_qs.filter(fecha_produccion__lte=dg_filters["end_date"])
    dashboard = _plan_status_dashboard(
        planes_qs,
        start_date=dg_filters["start_date"],
        end_date=dg_filters["end_date"],
        group_by=dg_filters["group_by"],
        limit=24,
    )
    planes_abiertos = planes_corte_qs.exclude(estado=PlanProduccion.ESTADO_CERRADO)[:20]
    planes_cerrados = planes_corte_qs.filter(estado=PlanProduccion.ESTADO_CERRADO)[:20]
    return render(
        request,
        "recetas/plan_produccion_dg_dashboard.html",
        {
            "plan_status_dashboard": dashboard,
            "dg_start_date": dg_filters["start_date"].isoformat() if dg_filters["start_date"] else "",
            "dg_end_date": dg_filters["end_date"].isoformat() if dg_filters["end_date"] else "",
            "dg_group_by": dg_filters["group_by"],
            "planes_abiertos": planes_abiertos,
            "planes_cerrados": planes_cerrados,
        },
    )


def _point_official_sales_stage_max_date() -> date | None:
    return PointDailySale.objects.filter(source_endpoint=OFFICIAL_POINT_SOURCE).aggregate(max_date=Max("sale_date")).get("max_date")


def _point_recent_sales_stage_max_date() -> date | None:
    return PointDailySale.objects.filter(source_endpoint=RECENT_POINT_SOURCE).aggregate(max_date=Max("sale_date")).get("max_date")


def _point_operational_sales_filters(*, start_date: date, end_date: date) -> Q:
    official_max = _point_official_sales_stage_max_date()
    q = Q()
    if official_max:
        official_end = min(end_date, official_max)
        if start_date <= official_end:
            q |= Q(source_endpoint=OFFICIAL_POINT_SOURCE, sale_date__gte=start_date, sale_date__lte=official_end)
        recent_start = max(start_date, official_max + timedelta(days=1))
    else:
        recent_start = start_date
    if recent_start <= end_date:
        q |= Q(source_endpoint=RECENT_POINT_SOURCE, sale_date__gte=recent_start, sale_date__lte=end_date)
    return q


def _point_operational_sales_rows_for_date(target_date: date):
    if PointDailySale.objects.filter(sale_date=target_date, source_endpoint=OFFICIAL_POINT_SOURCE).exists():
        return PointDailySale.objects.filter(sale_date=target_date, source_endpoint=OFFICIAL_POINT_SOURCE)
    return PointDailySale.objects.filter(sale_date=target_date, source_endpoint=RECENT_POINT_SOURCE)


def _point_operational_sales_history_queryset():
    official_max = _point_official_sales_stage_max_date()
    q = Q(source_endpoint=OFFICIAL_POINT_SOURCE)
    if official_max:
        q |= Q(source_endpoint=RECENT_POINT_SOURCE, sale_date__gt=official_max)
    else:
        q = Q(source_endpoint=RECENT_POINT_SOURCE)
    return PointDailySale.objects.filter(q)


def _latest_point_operational_cutoff_date() -> date:
    candidates: list[date] = []
    raw_candidates = [
        max([value for value in [_point_official_sales_stage_max_date(), _point_recent_sales_stage_max_date()] if value], default=None),
        PointDailyBranchIndicator.objects.aggregate(max_date=Max("indicator_date")).get("max_date"),
        PointProductionLine.objects.aggregate(max_date=Max("production_date")).get("max_date"),
        PointWasteLine.objects.aggregate(max_date=Max("movement_at")).get("max_date"),
        PointTransferLine.objects.filter(is_received=True).aggregate(max_date=Max("received_at")).get("max_date"),
        PointInventorySnapshot.objects.aggregate(max_date=Max("captured_at")).get("max_date"),
    ]
    for value in raw_candidates:
        if not value:
            continue
        if isinstance(value, datetime):
            candidates.append(timezone.localtime(value).date() if timezone.is_aware(value) else value.date())
        else:
            candidates.append(value)
    return min(candidates) if candidates else (timezone.localdate() - timedelta(days=1))


def _build_dg_operacion_dashboard_payload(
    *,
    start_date: date | None = None,
    end_date: date | None = None,
    group_by: str = "day",
    fecha_operacion: date | None = None,
) -> dict[str, Any]:
    if group_by not in {"day", "week", "month"}:
        group_by = "day"
    fecha_operacion = fecha_operacion or (_latest_point_operational_cutoff_date() + timedelta(days=1))
    planes_qs = PlanProduccion.objects.select_related("creado_por").prefetch_related("items").order_by("-fecha_produccion", "-id")
    plan_dashboard = _plan_status_dashboard(
        planes_qs,
        start_date=start_date,
        end_date=end_date,
        group_by=group_by,
        limit=18,
    )

    try:
        ventas_historicas_summary = _ventas_historicas_plan_summary()
        ventas_historicas_unavailable = False
    except (OperationalError, ProgrammingError):
        ventas_historicas_summary = {
            "available": False,
            "status": "Sin histórico",
            "tone": "warning",
            "detail": "La base diaria no está disponible en este entorno.",
            "date_label": "Sin cobertura",
            "active_days": 0,
            "expected_days": 0,
            "missing_days": 0,
            "branch_count": 0,
            "recipe_count": 0,
            "total_rows": 0,
            "total_units": Decimal("0"),
            "top_branches": [],
            "top_recipes": [],
        }
        ventas_historicas_unavailable = True

    sucursales = list(sucursales_operativas())
    resumen_cierre = _resumen_cierre_sucursales_reabasto(fecha_operacion, sucursales)
    demand_history_summary = _reabasto_demand_history_summary(fecha_operacion)
    reabasto_stage = (
        "Listo 8:00 AM"
        if resumen_cierre["listo_8am"]
        else "Con rezago"
        if (resumen_cierre["tardias"] or resumen_cierre["pendientes"])
        else "Sin captura"
    )
    reabasto_tone = "success" if resumen_cierre["listo_8am"] else "warning"
    reabasto_detail = (
        "Todas las sucursales enviaron en tiempo para el arranque de CEDIS."
        if resumen_cierre["listo_8am"]
        else f"Pendientes {resumen_cierre['pendientes']} · Tardías {resumen_cierre['tardias']} · Borrador {resumen_cierre['borrador']}."
    )
    point_exec_summary = _build_point_exec_summary(fecha_operacion)
    point_waste_summary = _build_point_waste_summary(fecha_operacion)
    point_central_flow_summary = _build_point_central_flow_summary(fecha_operacion)
    point_closure_summary = _build_point_closure_summary(fecha_operacion)
    point_chart_bundle = _build_point_chart_bundle(
        fecha_operacion=fecha_operacion,
        point_closure_summary=point_closure_summary,
    )

    return {
        "fecha_operacion": fecha_operacion,
        "plan_status_dashboard": plan_dashboard,
        "dg_start_date": start_date.isoformat() if start_date else "",
        "dg_end_date": end_date.isoformat() if end_date else "",
        "dg_group_by": group_by,
        "ventas_historicas_summary": ventas_historicas_summary,
        "ventas_historicas_unavailable": ventas_historicas_unavailable,
        "resumen_cierre": resumen_cierre,
        "reabasto_stage": reabasto_stage,
        "reabasto_tone": reabasto_tone,
        "reabasto_detail": reabasto_detail,
        "demand_history_summary": demand_history_summary,
        "point_exec_summary": point_exec_summary,
        "point_waste_summary": point_waste_summary,
        "point_central_flow_summary": point_central_flow_summary,
        "point_closure_summary": point_closure_summary,
        "point_chart_bundle": point_chart_bundle,
    }


def _point_operational_date_from_timestamp(value: datetime | None) -> date | None:
    if value is None:
        return None
    if timezone.is_aware(value):
        return timezone.localtime(value).date()
    return timezone.make_aware(value, timezone=timezone.get_current_timezone()).date()


def _point_waste_branch_label(row: PointWasteLine) -> tuple[str, str]:
    if row.erp_branch_id:
        return row.erp_branch.codigo, row.erp_branch.nombre
    external_id = (getattr(row.branch, "external_id", "") or "").strip()
    branch_name = (getattr(row.branch, "name", "") or "").strip()
    code = external_id or normalizar_nombre(branch_name).replace(" ", "_").upper() or "SIN_SUCURSAL"
    return code, branch_name or code


def _build_point_exec_summary(fecha_operacion: date) -> dict[str, Any]:
    cierre_fecha = fecha_operacion - timedelta(days=1)
    try:
        sales_qs = _point_operational_sales_rows_for_date(cierre_fecha).filter(
            branch__erp_branch_id__isnull=False,
            branch__erp_branch__activa=True,
        )
        indicator_qs = PointDailyBranchIndicator.objects.filter(
            indicator_date=cierre_fecha,
            branch__erp_branch_id__isnull=False,
            branch__erp_branch__activa=True,
        )
        month_sales_qs = PointDailySale.objects.filter(
            sale_date__year=cierre_fecha.year,
            sale_date__month=cierre_fecha.month,
            branch__erp_branch_id__isnull=False,
            branch__erp_branch__activa=True,
        ).filter(
            _point_operational_sales_filters(
                start_date=date(cierre_fecha.year, cierre_fecha.month, 1),
                end_date=cierre_fecha,
            )
        )
        month_indicator_qs = PointDailyBranchIndicator.objects.filter(
            indicator_date__year=cierre_fecha.year,
            indicator_date__month=cierre_fecha.month,
            branch__erp_branch_id__isnull=False,
            branch__erp_branch__activa=True,
        )

        latest_sales_amount = Decimal(str(sales_qs.aggregate(total=Sum("total_amount")).get("total") or 0))
        latest_units = Decimal(str(sales_qs.aggregate(total=Sum("quantity")).get("total") or 0))
        latest_tickets = int(indicator_qs.aggregate(total=Sum("total_tickets")).get("total") or 0)
        latest_indicator_amount = Decimal(str(indicator_qs.aggregate(total=Sum("total_amount")).get("total") or 0))
        latest_avg_ticket = (
            (latest_indicator_amount or latest_sales_amount) / Decimal(str(latest_tickets))
            if latest_tickets > 0
            else Decimal("0")
        )

        month_sales_amount, month_units = _partial_month_amount_quantity(
            start_date=date(cierre_fecha.year, cierre_fecha.month, 1),
            end_date=cierre_fecha,
        )
        month_tickets = int(month_indicator_qs.aggregate(total=Sum("total_tickets")).get("total") or 0)
        month_indicator_amount = Decimal(str(month_indicator_qs.aggregate(total=Sum("total_amount")).get("total") or 0))
        month_avg_ticket = (
            (month_indicator_amount or month_sales_amount) / Decimal(str(month_tickets))
            if month_tickets > 0
            else Decimal("0")
        )

        active_branch_count = sucursales_operativas().count()
        ticket_branch_count = indicator_qs.values("branch__erp_branch_id").distinct().count()
        top_branches = list(
            indicator_qs.values("branch__erp_branch__codigo", "branch__erp_branch__nombre")
            .annotate(total_amount=Sum("total_amount"), total_tickets=Sum("total_tickets"))
            .order_by("-total_amount", "branch__erp_branch__codigo")[:5]
        )
        for row in top_branches:
            total_tickets = int(row.get("total_tickets") or 0)
            total_amount = Decimal(str(row.get("total_amount") or 0))
            row["avg_ticket"] = (total_amount / Decimal(str(total_tickets))) if total_tickets > 0 else Decimal("0")

        return {
            "available": True,
            "closure_date": cierre_fecha,
            "latest_sales_amount": latest_sales_amount,
            "latest_units": latest_units,
            "latest_tickets": latest_tickets,
            "latest_avg_ticket": latest_avg_ticket,
            "month_sales_amount": month_sales_amount,
            "month_units": month_units,
            "month_tickets": month_tickets,
            "month_avg_ticket": month_avg_ticket,
            "ticket_data_available": latest_tickets > 0,
            "ticket_branch_count": ticket_branch_count,
            "active_branch_count": active_branch_count,
            "top_branches": top_branches,
            "source_label": "Point directo",
        }
    except (OperationalError, ProgrammingError):
        return {
            "available": False,
            "closure_date": cierre_fecha,
            "latest_sales_amount": Decimal("0"),
            "latest_units": Decimal("0"),
            "latest_tickets": 0,
            "latest_avg_ticket": Decimal("0"),
            "month_sales_amount": Decimal("0"),
            "month_units": Decimal("0"),
            "month_tickets": 0,
            "month_avg_ticket": Decimal("0"),
            "ticket_data_available": False,
            "ticket_branch_count": 0,
            "active_branch_count": 0,
            "top_branches": [],
            "source_label": "Sin fuente",
        }


def _build_point_waste_summary(fecha_operacion: date) -> dict[str, Any]:
    cierre_fecha = fecha_operacion - timedelta(days=1)
    try:
        window_start = timezone.make_aware(datetime.combine(cierre_fecha, datetime.min.time()), timezone.get_current_timezone())
        window_end = window_start + timedelta(days=2)
        waste_rows = [
            row
            for row in PointWasteLine.objects.filter(
                movement_at__gte=window_start,
                movement_at__lt=window_end,
            )
            .select_related("branch", "erp_branch")
            .order_by("movement_at", "branch__name", "item_name")
            if _point_operational_date_from_timestamp(row.movement_at) == cierre_fecha
        ]
        total_rows = len(waste_rows)
        total_qty = sum((Decimal(str(row.quantity or 0)) for row in waste_rows), Decimal("0"))
        total_cost = sum((Decimal(str(row.total_cost or 0)) for row in waste_rows), Decimal("0"))
        branch_buckets: dict[tuple[str, str], dict[str, Any]] = {}
        responsible_buckets: dict[str, dict[str, Any]] = {}
        for row in waste_rows:
            branch_code, branch_name = _point_waste_branch_label(row)
            branch_bucket = branch_buckets.setdefault(
                (branch_code, branch_name),
                {
                    "branch_code": branch_code,
                    "branch_name": branch_name,
                    "total_qty": Decimal("0"),
                    "total_cost": Decimal("0"),
                    "total_rows": 0,
                },
            )
            branch_bucket["total_qty"] += Decimal(str(row.quantity or 0))
            branch_bucket["total_cost"] += Decimal(str(row.total_cost or 0))
            branch_bucket["total_rows"] += 1

            responsible = (row.responsible or "").strip()
            if responsible:
                responsible_bucket = responsible_buckets.setdefault(
                    responsible,
                    {
                        "responsable_texto": responsible,
                        "total_qty": Decimal("0"),
                        "total_cost": Decimal("0"),
                        "total_rows": 0,
                    },
                )
                responsible_bucket["total_qty"] += Decimal(str(row.quantity or 0))
                responsible_bucket["total_cost"] += Decimal(str(row.total_cost or 0))
                responsible_bucket["total_rows"] += 1

        top_branches = sorted(
            branch_buckets.values(),
            key=lambda item: (-Decimal(str(item["total_qty"])), -Decimal(str(item["total_cost"])), str(item["branch_code"])),
        )[:5]
        top_responsibles = sorted(
            responsible_buckets.values(),
            key=lambda item: (-Decimal(str(item["total_qty"])), -Decimal(str(item["total_cost"])), str(item["responsable_texto"])),
        )[:5]
        branch_count = len(branch_buckets)
        responsible_count = len(responsible_buckets)
        if total_rows == 0:
            return {
                "available": True,
                "closure_date": cierre_fecha,
                "status": "Sin mermas registradas",
                "tone": "success",
                "detail": "Point no reportó mermas para el día operativo seleccionado.",
                "total_rows": 0,
                "total_qty": Decimal("0"),
                "total_cost": Decimal("0"),
                "branch_count": 0,
                "responsible_count": 0,
                "top_branches": [],
                "top_responsibles": [],
                "source_label": "Point directo",
            }
        return {
            "available": True,
            "closure_date": cierre_fecha,
            "status": "Mermas registradas",
            "tone": "warning" if total_qty > 0 else "success",
            "detail": f"{total_rows} registro(s) de merma capturados directamente en Point.",
            "total_rows": total_rows,
            "total_qty": total_qty,
            "total_cost": total_cost,
            "branch_count": branch_count,
            "responsible_count": responsible_count,
            "top_branches": top_branches,
            "top_responsibles": top_responsibles,
            "source_label": "Point directo",
        }
    except (OperationalError, ProgrammingError):
        return {
            "available": False,
            "closure_date": cierre_fecha,
            "status": "Sin mermas Point",
            "tone": "warning",
            "detail": "La base de mermas Point no está disponible en este entorno.",
            "total_rows": 0,
            "total_qty": Decimal("0"),
            "total_cost": Decimal("0"),
            "branch_count": 0,
            "responsible_count": 0,
            "top_branches": [],
            "top_responsibles": [],
            "source_label": "Sin fuente",
        }


def _build_point_central_flow_summary(fecha_operacion: date) -> dict[str, Any]:
    cierre_fecha = fecha_operacion - timedelta(days=1)
    settings = load_point_bridge_settings()
    allowed_production = {normalizar_nombre(value) for value in settings.production_storage_branches if value}
    allowed_transfer = {normalizar_nombre(value) for value in settings.transfer_storage_branches if value}
    try:
        production_qs = PointProductionLine.objects.filter(production_date=cierre_fecha).select_related("erp_branch", "branch")
        production_rows = [
            row for row in production_qs
            if normalizar_nombre(getattr(row.erp_branch, "nombre", "") or row.branch.name) in allowed_production
            and (row.erp_branch_id is None or getattr(row.erp_branch, "activa", False))
        ]
        transfer_qs = PointTransferLine.objects.filter(
            is_received=True,
            received_at__date=cierre_fecha,
        ).select_related("erp_destination_branch", "destination_branch", "origin_branch")
        transfer_rows = [
            row for row in transfer_qs
            if normalizar_nombre(getattr(row.erp_destination_branch, "nombre", "") or row.destination_branch.name) in allowed_transfer
            and (row.erp_destination_branch_id is None or getattr(row.erp_destination_branch, "activa", False))
        ]
        production_total_qty = sum((Decimal(str(row.produced_quantity or 0)) for row in production_rows), Decimal("0"))
        transfer_total_qty = sum((Decimal(str(row.received_quantity or 0)) for row in transfer_rows), Decimal("0"))
        production_item_count = len({((row.item_code or "").strip(), row.item_name.strip()) for row in production_rows})
        transfer_item_count = len({((row.item_code or "").strip(), row.item_name.strip()) for row in transfer_rows})
        top_transfer_origins: dict[str, dict[str, Any]] = {}
        for row in transfer_rows:
            label = f"{row.origin_branch.external_id} · {row.origin_branch.name}"
            bucket = top_transfer_origins.setdefault(label, {"origin": label, "total_qty": Decimal("0"), "total_rows": 0})
            bucket["total_qty"] += Decimal(str(row.received_quantity or 0))
            bucket["total_rows"] += 1
        top_transfer_origins_rows = sorted(
            top_transfer_origins.values(),
            key=lambda item: (-item["total_qty"], item["origin"]),
        )[:5]
        status = "Sin flujo central"
        tone = "warning"
        detail = "No hubo entradas directas ni transferencias recibidas hacia inventario central en el corte."
        if production_rows or transfer_rows:
            status = "Flujo central registrado"
            tone = "success"
            detail = (
                f"Producción directa {len(production_rows)} línea(s) y transferencias recibidas {len(transfer_rows)} línea(s) "
                "hacia inventario central."
            )
        return {
            "available": True,
            "closure_date": cierre_fecha,
            "status": status,
            "tone": tone,
            "detail": detail,
            "production_rows": len(production_rows),
            "production_total_qty": production_total_qty,
            "production_item_count": production_item_count,
            "transfer_rows": len(transfer_rows),
            "transfer_total_qty": transfer_total_qty,
            "transfer_item_count": transfer_item_count,
            "top_transfer_origins": top_transfer_origins_rows,
        }
    except (OperationalError, ProgrammingError):
        return {
            "available": False,
            "closure_date": cierre_fecha,
            "status": "Sin flujo central Point",
            "tone": "warning",
            "detail": "La base Point de producción/transferencias no está disponible en este entorno.",
            "production_rows": 0,
            "production_total_qty": Decimal("0"),
            "production_item_count": 0,
            "transfer_rows": 0,
            "transfer_total_qty": Decimal("0"),
            "transfer_item_count": 0,
            "top_transfer_origins": [],
        }


def _build_point_closure_summary(fecha_operacion: date) -> dict[str, Any]:
    try:
        cierre_fecha = fecha_operacion - timedelta(days=1)
        tz = timezone.get_current_timezone()
        apertura_dt = timezone.make_aware(datetime.combine(cierre_fecha, datetime.min.time()), tz)
        siguiente_dt = apertura_dt + timedelta(days=1)
        corte_cierre_dt = siguiente_dt + timedelta(hours=12)

        point_branches = list(
            _point_operational_sales_history_queryset().filter(
                branch__erp_branch_id__isnull=False,
                branch__erp_branch__activa=True,
            )
            .values_list("branch__erp_branch_id", "branch__erp_branch__codigo", "branch__erp_branch__nombre")
            .distinct()
        )
        branch_map: dict[int, dict[str, Any]] = {
            branch_id: {"branch_id": branch_id, "codigo": codigo, "nombre": nombre}
            for branch_id, codigo, nombre in point_branches
        }
        for branch in sucursales_operativas().only("id", "codigo", "nombre"):
            branch_map.setdefault(branch.id, {"branch_id": branch.id, "codigo": branch.codigo, "nombre": branch.nombre})

        produced_map: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
        for row in (
            PointProductionLine.objects.filter(
                production_date=cierre_fecha,
                is_insumo=False,
                receta_id__isnull=False,
                erp_branch_id__isnull=False,
                erp_branch__activa=True,
            )
            .values("erp_branch_id")
            .annotate(total=Sum("produced_quantity"))
        ):
            produced_map[int(row["erp_branch_id"])] += Decimal(str(row.get("total") or 0))

        transfer_in_map: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
        for row in (
            PointTransferLine.objects.filter(
                is_insumo=False,
                receta_id__isnull=False,
                is_received=True,
                received_at__date=cierre_fecha,
                erp_destination_branch_id__isnull=False,
                erp_destination_branch__activa=True,
            )
            .values("erp_destination_branch_id")
            .annotate(total=Sum("received_quantity"))
        ):
            transfer_in_map[int(row["erp_destination_branch_id"])] += Decimal(str(row.get("total") or 0))

        sales_map: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
        tickets_map: dict[int, int] = defaultdict(int)
        sales_amount_map: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
        indicator_tickets_map: dict[int, int] = defaultdict(int)
        indicator_sales_amount_map: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
        for row in (
            _point_operational_sales_rows_for_date(cierre_fecha).filter(
                branch__erp_branch_id__isnull=False,
                branch__erp_branch__activa=True,
            )
            .values("branch__erp_branch_id")
            .annotate(total=Sum("quantity"), total_tickets=Sum("tickets"), total_sales=Sum("total_amount"))
        ):
            sales_map[int(row["branch__erp_branch_id"])] += Decimal(str(row.get("total") or 0))
            tickets_map[int(row["branch__erp_branch_id"])] += int(row.get("total_tickets") or 0)
            sales_amount_map[int(row["branch__erp_branch_id"])] += Decimal(str(row.get("total_sales") or 0))

        for row in (
            PointDailyBranchIndicator.objects.filter(
                indicator_date=cierre_fecha,
                branch__erp_branch_id__isnull=False,
                branch__erp_branch__activa=True,
            )
            .values("branch__erp_branch_id")
            .annotate(total_tickets=Sum("total_tickets"), total_sales=Sum("total_amount"))
        ):
            indicator_tickets_map[int(row["branch__erp_branch_id"])] += int(row.get("total_tickets") or 0)
            indicator_sales_amount_map[int(row["branch__erp_branch_id"])] += Decimal(str(row.get("total_sales") or 0))

        waste_map: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
        for row in (
            MermaPOS.objects.filter(
                fecha=cierre_fecha,
                sucursal_id__isnull=False,
                sucursal__activa=True,
            )
            .values("sucursal_id")
            .annotate(total=Sum("cantidad"))
        ):
            waste_map[int(row["sucursal_id"])] += Decimal(str(row.get("total") or 0))

        opening_rows = (
            PointInventorySnapshot.objects.filter(
                captured_at__lt=apertura_dt,
                branch__erp_branch_id__isnull=False,
                branch__erp_branch__activa=True,
            )
            .select_related("branch__erp_branch")
            .order_by("branch_id", "product_id", "-captured_at")
        )
        opening_map: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
        seen_opening: set[tuple[int, int]] = set()
        for snapshot in opening_rows:
            key = (snapshot.branch_id, snapshot.product_id)
            if key in seen_opening:
                continue
            seen_opening.add(key)
            if snapshot.branch.erp_branch_id:
                opening_map[int(snapshot.branch.erp_branch_id)] += Decimal(str(snapshot.stock or 0))

        closing_rows = (
            PointInventorySnapshot.objects.filter(
                captured_at__lt=corte_cierre_dt,
                branch__erp_branch_id__isnull=False,
                branch__erp_branch__activa=True,
            )
            .select_related("branch__erp_branch")
            .order_by("branch_id", "product_id", "-captured_at")
        )
        closing_map: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
        closing_snapshot_recent: set[int] = set()
        seen_closing: set[tuple[int, int]] = set()
        for snapshot in closing_rows:
            key = (snapshot.branch_id, snapshot.product_id)
            if key in seen_closing:
                continue
            seen_closing.add(key)
            if snapshot.branch.erp_branch_id:
                branch_id = int(snapshot.branch.erp_branch_id)
                closing_map[branch_id] += Decimal(str(snapshot.stock or 0))
                if snapshot.captured_at >= siguiente_dt:
                    closing_snapshot_recent.add(branch_id)

        rows: list[dict[str, Any]] = []
        total_cuadra = 0
        total_minor = 0
        total_critical = 0
        total_no_close = 0

        for branch_id, branch in sorted(branch_map.items(), key=lambda item: (item[1]["codigo"], item[1]["nombre"])):
            opening_units = opening_map.get(branch_id, Decimal("0"))
            produced_units = produced_map.get(branch_id, Decimal("0"))
            transfer_units = transfer_in_map.get(branch_id, Decimal("0"))
            sold_units = sales_map.get(branch_id, Decimal("0"))
            sold_tickets = int(indicator_tickets_map.get(branch_id) or tickets_map.get(branch_id, 0))
            sold_amount = indicator_sales_amount_map.get(branch_id) or sales_amount_map.get(branch_id, Decimal("0"))
            avg_ticket = (sold_amount / Decimal(str(sold_tickets))) if sold_tickets > 0 else Decimal("0")
            waste_units = waste_map.get(branch_id, Decimal("0"))
            expected_units = opening_units + produced_units + transfer_units - sold_units - waste_units
            closing_units = closing_map.get(branch_id, Decimal("0"))
            variance_units = closing_units - expected_units
            has_flow = any(
                value != Decimal("0")
                for value in (opening_units, produced_units, transfer_units, sold_units, waste_units, closing_units)
            )
            has_closing = branch_id in closing_snapshot_recent
            if not has_flow:
                status = "Sin movimiento"
                tone = "neutral"
            elif not has_closing:
                status = "Sin cierre"
                tone = "warning"
                total_no_close += 1
            else:
                abs_variance = abs(variance_units)
                if abs_variance <= Decimal("3"):
                    status = "Cuadra"
                    tone = "success"
                    total_cuadra += 1
                elif abs_variance <= Decimal("10"):
                    status = "Desviacion menor"
                    tone = "warning"
                    total_minor += 1
                else:
                    status = "Desviacion critica"
                    tone = "danger"
                    total_critical += 1
            rows.append(
                {
                    "branch_id": branch_id,
                    "branch_label": f"{branch['codigo']} · {branch['nombre']}",
                    "opening_units": opening_units,
                    "produced_units": produced_units,
                    "transfer_units": transfer_units,
                    "sold_units": sold_units,
                    "sold_tickets": sold_tickets,
                    "avg_ticket": avg_ticket,
                    "waste_units": waste_units,
                    "expected_units": expected_units,
                    "closing_units": closing_units,
                    "variance_units": variance_units,
                    "has_closing_snapshot": has_closing,
                    "status": status,
                    "tone": tone,
                }
            )

        rows = [row for row in rows if any(row[key] != Decimal("0") for key in ("produced_units", "transfer_units", "sold_units", "waste_units", "closing_units", "opening_units"))]
        rows.sort(key=lambda row: (0 if row["tone"] == "danger" else 1 if row["tone"] == "warning" else 2, -abs(row["variance_units"])))

        if total_critical > 0:
            status = "Desviaciones criticas"
            tone = "danger"
            detail = f"{total_critical} sucursal(es) no cuadran contra snapshot de cierre."
        elif total_minor > 0 or total_no_close > 0:
            status = "Pendiente de cuadre"
            tone = "warning"
            detail = f"Sin cierre {total_no_close} · Desviacion menor {total_minor}."
        else:
            status = "Cuadre controlado"
            tone = "success"
            detail = "Las sucursales con cierre disponible quedan dentro de tolerancia."

        return {
            "available": True,
            "closure_date": cierre_fecha,
            "status": status,
            "tone": tone,
            "detail": detail,
            "branch_count": len(rows),
            "cuadra": total_cuadra,
            "minor": total_minor,
            "critical": total_critical,
            "without_closure": total_no_close,
            "rows": rows[:25],
        }
    except (OperationalError, ProgrammingError):
        return {
            "available": False,
            "closure_date": fecha_operacion - timedelta(days=1),
            "status": "Sin cuadre Point",
            "tone": "warning",
            "detail": "La base Point no está disponible en este entorno.",
            "branch_count": 0,
            "cuadra": 0,
            "minor": 0,
            "critical": 0,
            "without_closure": 0,
            "rows": [],
        }


def _point_week_start(target_date: date) -> date:
    return target_date - timedelta(days=target_date.weekday())


_POINT_DASHBOARD_CATEGORY_ORDER = [
    "Bollo",
    "Galletas",
    "Pastel Chico",
    "Pastel Mediano",
    "Pastel Grande",
    "Pastel Mini",
    "Rebanada",
    "Individual",
    "Pay Mediano",
    "Pay Grande",
    "Empanadas",
    "Vasos Chico",
    "Vasos Mediano",
    "Vasos Grande",
    "Accesorios",
    "Sin categoría",
]


def _point_operational_category(*, category: str = "", family: str = "", item_name: str = "") -> str:
    category_clean = (category or "").strip()
    family_clean = (family or "").strip()
    item_clean = (item_name or "").strip()
    norm = normalizar_nombre(" ".join([category_clean, family_clean, item_clean]).strip())

    if not norm:
        return "Sin categoría"
    if any(token in norm for token in ("servicio domicilio", "vela", "tarjeta regalo", "accesorio", "alegria", "pirotecnia", "espagueti")):
        return "Accesorios"
    if "rebanada" in norm or norm.endswith(" r") or " reb" in norm:
        return "Rebanada"
    if "vaso" in norm:
        if "grande" in norm:
            return "Vasos Grande"
        if "mediano" in norm:
            return "Vasos Mediano"
        if "chico" in norm:
            return "Vasos Chico"
        return "Vasos Mediano"
    if "empanada" in norm:
        return "Empanadas"
    if "galleta" in norm or "bolitas de nuez" in norm:
        return "Galletas"
    if "bollo" in norm:
        return "Bollo"
    if "cheesecake" in norm or "individual" in norm:
        return "Individual"
    if "pastel" in norm:
        if "mini" in norm:
            return "Pastel Mini"
        if "grande" in norm:
            return "Pastel Grande"
        if "mediano" in norm:
            return "Pastel Mediano"
        if "chico" in norm:
            return "Pastel Chico"
        return "Pastel Mediano"
    if "pay" in norm or "pie" in norm:
        if "grande" in norm:
            return "Pay Grande"
        if "mediano" in norm:
            return "Pay Mediano"
        return "Pay Mediano"

    normalized_category = normalizar_nombre(category_clean)
    normalized_family = normalizar_nombre(family_clean)
    family_category_map = {
        "bollo": "Bollo",
        "galletas": "Galletas",
        "empanadas": "Empanadas",
        "cheesecakes": "Individual",
        "pastel": "Pastel Mediano",
        "pay": "Pay Mediano",
    }
    if normalized_category in {normalizar_nombre(label) for label in _POINT_DASHBOARD_CATEGORY_ORDER}:
        for label in _POINT_DASHBOARD_CATEGORY_ORDER:
            if normalizar_nombre(label) == normalized_category:
                return label
    if normalized_family in family_category_map:
        return family_category_map[normalized_family]
    return "Sin categoría"


def _build_point_chart_bundle(
    *,
    fecha_operacion: date,
    point_closure_summary: dict[str, Any],
) -> dict[str, Any]:
    cierre_fecha = fecha_operacion - timedelta(days=1)
    current_week_start = _point_week_start(cierre_fecha)
    week_starts = [current_week_start - timedelta(weeks=offset) for offset in range(5, -1, -1)]
    weekly_labels = [f"{week_start:%d %b}" for week_start in week_starts]
    weekly_sales_amount = {week_start: Decimal("0") for week_start in week_starts}
    weekly_sales_units = {week_start: Decimal("0") for week_start in week_starts}
    weekly_waste_units = {week_start: Decimal("0") for week_start in week_starts}
    weekly_production_units = {week_start: Decimal("0") for week_start in week_starts}
    weekly_transfer_units = {week_start: Decimal("0") for week_start in week_starts}
    week_lookup = set(week_starts)

    try:
        for row in (
            PointDailySale.objects.filter(
                sale_date__gte=week_starts[0],
                sale_date__lte=cierre_fecha,
                branch__erp_branch_id__isnull=False,
                branch__erp_branch__activa=True,
            ).filter(
                _point_operational_sales_filters(start_date=week_starts[0], end_date=cierre_fecha)
            )
            .values("sale_date")
            .annotate(total_sales=Sum("total_amount"), total_units=Sum("quantity"))
        ):
            week_start = _point_week_start(row["sale_date"])
            if week_start not in week_lookup:
                continue
            weekly_sales_amount[week_start] += Decimal(str(row.get("total_sales") or 0))
            weekly_sales_units[week_start] += Decimal(str(row.get("total_units") or 0))

        waste_window_start = timezone.make_aware(
            datetime.combine(week_starts[0], datetime.min.time()),
            timezone.get_current_timezone(),
        )
        waste_window_end = timezone.make_aware(
            datetime.combine(cierre_fecha + timedelta(days=2), datetime.min.time()),
            timezone.get_current_timezone(),
        )
        waste_rows_window = list(
            PointWasteLine.objects.filter(
                movement_at__gte=waste_window_start,
                movement_at__lt=waste_window_end,
            ).select_related("branch", "erp_branch")
        )
        waste_rows_by_day: dict[date, list[PointWasteLine]] = defaultdict(list)
        for row in waste_rows_window:
            operational_date = _point_operational_date_from_timestamp(row.movement_at)
            if operational_date is None or operational_date < week_starts[0] or operational_date > cierre_fecha:
                continue
            waste_rows_by_day[operational_date].append(row)

        for operational_date, waste_rows in waste_rows_by_day.items():
            week_start = _point_week_start(operational_date)
            if week_start not in week_lookup:
                continue
            weekly_waste_units[week_start] += sum(
                (Decimal(str(row.quantity or 0)) for row in waste_rows),
                Decimal("0"),
            )

        settings = load_point_bridge_settings()
        production_branch_names = {normalizar_nombre(value) for value in settings.production_storage_branches if str(value).strip()}
        transfer_branch_names = {normalizar_nombre(value) for value in settings.transfer_storage_branches if str(value).strip()}

        for row in PointProductionLine.objects.filter(
            production_date__gte=week_starts[0],
            production_date__lte=cierre_fecha,
        ).select_related("erp_branch", "branch"):
            branch_name = normalizar_nombre(getattr(row.erp_branch, "nombre", "") or row.branch.name)
            if branch_name not in production_branch_names:
                continue
            if row.erp_branch_id is not None and not getattr(row.erp_branch, "activa", False):
                continue
            week_start = _point_week_start(row.production_date)
            if week_start not in week_lookup:
                continue
            weekly_production_units[week_start] += Decimal(str(row.produced_quantity or 0))

        for row in PointTransferLine.objects.filter(
            is_received=True,
            received_at__date__gte=week_starts[0],
            received_at__date__lte=cierre_fecha,
        ).select_related("erp_destination_branch", "destination_branch"):
            branch_name = normalizar_nombre(getattr(row.erp_destination_branch, "nombre", "") or row.destination_branch.name)
            if branch_name not in transfer_branch_names:
                continue
            if row.erp_destination_branch_id is not None and not getattr(row.erp_destination_branch, "activa", False):
                continue
            received_date = row.received_at.date() if row.received_at else None
            if not received_date:
                continue
            week_start = _point_week_start(received_date)
            if week_start not in week_lookup:
                continue
            weekly_transfer_units[week_start] += Decimal(str(row.received_quantity or 0))

        waste_share_buckets: dict[tuple[str, str], Decimal] = defaultdict(lambda: Decimal("0"))
        for row in waste_rows_by_day.get(cierre_fecha, []):
            branch_code, branch_name = _point_waste_branch_label(row)
            waste_share_buckets[(branch_code, branch_name)] += Decimal(str(row.quantity or 0))
        waste_share_rows = [
            {"branch_code": branch_code, "branch_name": branch_name, "total_qty": total_qty}
            for (branch_code, branch_name), total_qty in sorted(
                waste_share_buckets.items(),
                key=lambda item: (-item[1], item[0][0]),
            )[:8]
        ]

        current_week_start = _point_week_start(cierre_fecha)
        sales_categories: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
        production_categories: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))

        for row in (
            PointDailySale.objects.filter(
                sale_date__gte=current_week_start,
                sale_date__lte=cierre_fecha,
                branch__erp_branch_id__isnull=False,
                branch__erp_branch__activa=True,
            ).filter(
                _point_operational_sales_filters(start_date=current_week_start, end_date=cierre_fecha)
            )
            .select_related("product")
        ):
            category = _point_operational_category(
                category=getattr(row.product, "category", ""),
                family=str((row.product.metadata or {}).get("family") or ""),
                item_name=getattr(row.product, "name", ""),
            )
            sales_categories[category] += Decimal(str(row.quantity or 0))

        for row in PointProductionLine.objects.filter(
            production_date__gte=current_week_start,
            production_date__lte=cierre_fecha,
            is_insumo=False,
            receta_id__isnull=False,
        ).select_related("erp_branch", "branch", "receta"):
            branch_name = normalizar_nombre(getattr(row.erp_branch, "nombre", "") or row.branch.name)
            if branch_name not in production_branch_names:
                continue
            if row.erp_branch_id is not None and not getattr(row.erp_branch, "activa", False):
                continue
            category = _point_operational_category(
                category=getattr(row.receta, "categoria", ""),
                family=getattr(row.receta, "familia", ""),
                item_name=row.item_name,
            )
            production_categories[category] += Decimal(str(row.produced_quantity or 0))

        for row in PointTransferLine.objects.filter(
            is_received=True,
            received_at__date__gte=current_week_start,
            received_at__date__lte=cierre_fecha,
        ).select_related("erp_destination_branch", "destination_branch", "receta"):
            branch_name = normalizar_nombre(getattr(row.erp_destination_branch, "nombre", "") or row.destination_branch.name)
            if branch_name not in transfer_branch_names:
                continue
            if row.erp_destination_branch_id is not None and not getattr(row.erp_destination_branch, "activa", False):
                continue

        ranked_categories = sorted(
            set(sales_categories.keys()) | set(production_categories.keys()),
            key=lambda category: (
                _POINT_DASHBOARD_CATEGORY_ORDER.index(category) if category in _POINT_DASHBOARD_CATEGORY_ORDER else 999,
                -(sales_categories.get(category, Decimal("0")) + production_categories.get(category, Decimal("0"))),
                category,
            ),
        )
        ranked_categories = [
            category
            for category in ranked_categories
            if sales_categories.get(category, Decimal("0")) > 0 or production_categories.get(category, Decimal("0")) > 0
        ][:10]
        unclassified_sales_units = sales_categories.get("Sin categoría", Decimal("0"))
        unclassified_production_units = production_categories.get("Sin categoría", Decimal("0"))
        classified_sales_units = sum(
            value for key, value in sales_categories.items() if key != "Sin categoría"
        )
        classified_production_units = sum(
            value for key, value in production_categories.items() if key != "Sin categoría"
        )
        total_sales_category_units = classified_sales_units + unclassified_sales_units
        total_inflow_category_units = classified_production_units + unclassified_production_units
        sales_classification_pct = (
            (classified_sales_units / total_sales_category_units * Decimal("100"))
            if total_sales_category_units > 0
            else Decimal("100")
        )
        inflow_classification_pct = (
            (classified_production_units / total_inflow_category_units * Decimal("100"))
            if total_inflow_category_units > 0
            else Decimal("100")
        )

        branch_rows = list(point_closure_summary.get("rows") or [])
        branch_balance_rows = sorted(
            branch_rows,
            key=lambda row: (
                -(Decimal(str(row["sold_units"])) + Decimal(str(row["produced_units"])) + Decimal(str(row["transfer_units"]))),
                row["branch_label"],
            ),
        )[:8]
        avg_ticket_rows = [row for row in branch_rows if int(row.get("sold_tickets") or 0) > 0]
        avg_ticket_rows.sort(key=lambda row: (-Decimal(str(row["avg_ticket"])), row["branch_label"]))
        avg_ticket_rows = avg_ticket_rows[:8]
        ticket_branch_coverage_pct = (
            Decimal(str(len(avg_ticket_rows))) / Decimal(str(len(branch_rows))) * Decimal("100")
            if branch_rows
            else Decimal("0")
        )

        return {
            "available": True,
            "ticket_data_available": bool(avg_ticket_rows),
            "ticket_branch_coverage_pct": float(ticket_branch_coverage_pct),
            "ticket_branch_count": len(avg_ticket_rows),
            "weekly_sales_amount": {
                "labels": weekly_labels,
                "values": [float(weekly_sales_amount[week_start]) for week_start in week_starts],
            },
            "weekly_ops_units": {
                "labels": weekly_labels,
                "sales": [float(weekly_sales_units[week_start]) for week_start in week_starts],
                "waste": [float(weekly_waste_units[week_start]) for week_start in week_starts],
                "production": [float(weekly_production_units[week_start]) for week_start in week_starts],
                "transfers": [float(weekly_transfer_units[week_start]) for week_start in week_starts],
            },
            "waste_share": {
                "labels": [row["branch_code"] for row in waste_share_rows],
                "values": [float(Decimal(str(row.get("total_qty") or 0))) for row in waste_share_rows],
            },
            "ticket_ranking": {
                "labels": [row["branch_label"].split(" · ")[0] for row in avg_ticket_rows],
                "values": [float(Decimal(str(row["avg_ticket"]))) for row in avg_ticket_rows],
            },
            "branch_balance": {
                "labels": [row["branch_label"].split(" · ")[0] for row in branch_balance_rows],
                "sales": [float(Decimal(str(row["sold_units"]))) for row in branch_balance_rows],
                "inflow": [float(Decimal(str(row["produced_units"])) + Decimal(str(row["transfer_units"]))) for row in branch_balance_rows],
                "waste": [float(Decimal(str(row["waste_units"]))) for row in branch_balance_rows],
            },
            "variance_map": {
                "labels": [row["branch_label"].split(" · ")[0] for row in branch_rows],
                "values": [float(Decimal(str(row["variance_units"]))) for row in branch_rows],
            },
            "category_balance": {
                "labels": ranked_categories,
                "sales": [float(sales_categories.get(category, Decimal("0"))) for category in ranked_categories],
                "production": [float(production_categories.get(category, Decimal("0"))) for category in ranked_categories],
                "inflow_label": "Producción directa",
                "unclassified_sales_units": float(unclassified_sales_units),
                "unclassified_production_units": float(unclassified_production_units),
                "classified_sales_units": float(classified_sales_units),
                "classified_production_units": float(classified_production_units),
                "sales_classification_pct": float(sales_classification_pct),
                "inflow_classification_pct": float(inflow_classification_pct),
            },
        }
    except (OperationalError, ProgrammingError):
        return {
            "available": False,
            "ticket_data_available": False,
            "ticket_branch_coverage_pct": 0.0,
            "ticket_branch_count": 0,
            "weekly_sales_amount": {"labels": [], "values": []},
            "weekly_ops_units": {"labels": [], "sales": [], "waste": [], "production": [], "transfers": []},
            "waste_share": {"labels": [], "values": []},
            "ticket_ranking": {"labels": [], "values": []},
            "branch_balance": {"labels": [], "sales": [], "inflow": [], "waste": []},
            "variance_map": {"labels": [], "values": []},
            "category_balance": {
                "labels": [],
                "sales": [],
                "production": [],
                "inflow_label": "Producción directa",
                "unclassified_sales_units": 0.0,
                "unclassified_production_units": 0.0,
                "classified_sales_units": 0.0,
                "classified_production_units": 0.0,
                "sales_classification_pct": 0.0,
                "inflow_classification_pct": 0.0,
            },
        }


def _iso_week_label(week_start: date) -> str:
    iso_year, iso_week, _ = week_start.isocalendar()
    return f"{iso_year}-W{iso_week:02d}"


def _safe_delta_pct(current: Decimal, previous: Decimal) -> Decimal | None:
    if previous == 0:
        return None
    return ((current - previous) / previous * Decimal("100")).quantize(Decimal("0.1"))


def _build_produccion_cedis_weekly_dashboard(base_date: date | None = None) -> dict[str, Any]:
    base_date = base_date or (timezone.localdate() - timedelta(days=1))
    current_week_start = _point_week_start(base_date)
    current_week_end = current_week_start + timedelta(days=6)
    week_starts = [current_week_start - timedelta(weeks=offset) for offset in range(2, -1, -1)]
    week_lookup = set(week_starts)

    settings = load_point_bridge_settings()
    production_branch_names = {normalizar_nombre(value) for value in settings.production_storage_branches if str(value).strip()}
    transfer_branch_names = {normalizar_nombre(value) for value in settings.transfer_storage_branches if str(value).strip()}
    sales_excluded_branch_names = {normalizar_nombre(value) for value in settings.sales_excluded_branches if str(value).strip()}

    weekly: dict[date, dict[str, Any]] = {
        week_start: {
            "week_start": week_start,
            "week_end": week_start + timedelta(days=6),
            "week_label": _iso_week_label(week_start),
            "sales_amount": Decimal("0"),
            "sales_units": Decimal("0"),
            "tickets": 0,
            "production_units": Decimal("0"),
            "transfer_units": Decimal("0"),
            "waste_cost": Decimal("0"),
            "waste_units": Decimal("0"),
        }
        for week_start in week_starts
    }

    sales_by_product: dict[str, dict[str, Decimal]] = defaultdict(lambda: {"amount": Decimal("0"), "qty": Decimal("0")})
    waste_by_product: dict[str, dict[str, Decimal]] = defaultdict(lambda: {"amount": Decimal("0"), "qty": Decimal("0")})
    sales_by_category: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    production_by_category: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    waste_by_branch: dict[str, dict[str, Decimal]] = defaultdict(lambda: {"cost": Decimal("0"), "qty": Decimal("0")})
    next_week_categories: dict[str, dict[str, Decimal | list[Decimal]]] = defaultdict(
        lambda: {"sales_amount_values": [], "sales_qty_values": [], "production_values": []}
    )

    sales_rows = (
        PointDailySale.objects.filter(
            sale_date__gte=week_starts[0],
            sale_date__lte=current_week_end,
            branch__erp_branch_id__isnull=False,
            branch__erp_branch__activa=True,
        ).filter(
            _point_operational_sales_filters(start_date=week_starts[0], end_date=current_week_end)
        )
        .select_related("branch__erp_branch", "product")
    )
    for row in sales_rows:
        branch_name = normalizar_nombre(getattr(row.branch.erp_branch, "nombre", "") or row.branch.name)
        if branch_name in sales_excluded_branch_names:
            continue
        week_start = _point_week_start(row.sale_date)
        if week_start not in week_lookup:
            continue
        bucket = weekly[week_start]
        qty = Decimal(str(row.quantity or 0))
        amount = Decimal(str(row.total_amount or 0))
        bucket["sales_units"] += qty
        bucket["sales_amount"] += amount
        category = _point_operational_category(
            category=getattr(row.product, "category", ""),
            family=str((row.product.metadata or {}).get("family") or ""),
            item_name=getattr(row.product, "name", ""),
        )
        next_week_categories[category]["sales_amount_values"].append(amount)
        next_week_categories[category]["sales_qty_values"].append(qty)
        if week_start == current_week_start:
            sales_by_product[row.product.name] = {
                "amount": sales_by_product[row.product.name]["amount"] + amount,
                "qty": sales_by_product[row.product.name]["qty"] + qty,
            }
            sales_by_category[category] += qty

    indicator_rows = (
        PointDailyBranchIndicator.objects.filter(
            indicator_date__gte=week_starts[0],
            indicator_date__lte=current_week_end,
            branch__erp_branch_id__isnull=False,
            branch__erp_branch__activa=True,
        )
        .select_related("branch__erp_branch")
    )
    for row in indicator_rows:
        branch_name = normalizar_nombre(getattr(row.branch.erp_branch, "nombre", "") or row.branch.name)
        if branch_name in sales_excluded_branch_names:
            continue
        week_start = _point_week_start(row.indicator_date)
        if week_start not in week_lookup:
            continue
        bucket = weekly[week_start]
        bucket["tickets"] += int(row.total_tickets or 0)

    production_rows = (
        PointProductionLine.objects.filter(
            production_date__gte=week_starts[0],
            production_date__lte=current_week_end,
            is_insumo=False,
        )
        .select_related("erp_branch", "branch", "receta")
    )
    for row in production_rows:
        branch_name = normalizar_nombre(getattr(row.erp_branch, "nombre", "") or row.branch.name)
        if branch_name not in production_branch_names:
            continue
        if row.erp_branch_id is not None and not getattr(row.erp_branch, "activa", False):
            continue
        week_start = _point_week_start(row.production_date)
        if week_start not in week_lookup:
            continue
        qty = Decimal(str(row.produced_quantity or 0))
        weekly[week_start]["production_units"] += qty
        category = _point_operational_category(
            category=getattr(row.receta, "categoria", ""),
            family=getattr(row.receta, "familia", ""),
            item_name=row.item_name,
        )
        next_week_categories[category]["production_values"].append(qty)
        if week_start == current_week_start:
            production_by_category[category] += qty

    transfer_rows = (
        PointTransferLine.objects.filter(
            is_received=True,
            received_at__date__gte=week_starts[0],
            received_at__date__lte=current_week_end,
        )
        .select_related("erp_destination_branch", "destination_branch", "receta")
    )
    for row in transfer_rows:
        branch_name = normalizar_nombre(getattr(row.erp_destination_branch, "nombre", "") or row.destination_branch.name)
        if branch_name not in transfer_branch_names:
            continue
        if row.erp_destination_branch_id is not None and not getattr(row.erp_destination_branch, "activa", False):
            continue
        received_date = row.received_at.date() if row.received_at else None
        if not received_date:
            continue
        week_start = _point_week_start(received_date)
        if week_start not in week_lookup:
            continue
        qty = Decimal(str(row.received_quantity or 0))
        weekly[week_start]["transfer_units"] += qty

    waste_rows = (
        PointWasteLine.objects.filter(
            movement_at__date__gte=week_starts[0],
            movement_at__date__lte=current_week_end,
            erp_branch_id__isnull=False,
            erp_branch__activa=True,
        )
        .select_related("erp_branch")
    )
    for row in waste_rows:
        movement_date = row.movement_at.date()
        week_start = _point_week_start(movement_date)
        if week_start not in week_lookup:
            continue
        qty = Decimal(str(row.quantity or 0))
        amount = Decimal(str(row.total_cost or 0))
        weekly[week_start]["waste_units"] += qty
        weekly[week_start]["waste_cost"] += amount
        if week_start == current_week_start:
            waste_by_product[row.item_name] = {
                "amount": waste_by_product[row.item_name]["amount"] + amount,
                "qty": waste_by_product[row.item_name]["qty"] + qty,
            }
            branch_label = getattr(row.erp_branch, "codigo", "") or getattr(row.erp_branch, "nombre", "") or "Sin sucursal"
            waste_by_branch[branch_label]["cost"] += amount
            waste_by_branch[branch_label]["qty"] += qty

    for bucket in weekly.values():
        sales_units = bucket["sales_units"]
        tickets = int(bucket["tickets"])
        sales_amount = bucket["sales_amount"]
        inflow = bucket["production_units"] + bucket["transfer_units"]
        bucket["avg_ticket"] = (sales_amount / Decimal(str(tickets))) if tickets > 0 else Decimal("0")
        bucket["coverage_ratio"] = (inflow / sales_units) if sales_units > 0 else Decimal("0")
        bucket["inflow_units"] = inflow

    current = weekly[current_week_start]
    previous = weekly[week_starts[-2]]

    top_sales_rows = sorted(
        [
            {"product": name, "amount": values["amount"], "qty": values["qty"]}
            for name, values in sales_by_product.items()
        ],
        key=lambda row: (-row["amount"], row["product"]),
    )[:10]
    total_top_sales_amount = sum((row["amount"] for row in top_sales_rows), Decimal("0"))
    for row in top_sales_rows:
        row["share_pct"] = ((row["amount"] / total_top_sales_amount) * Decimal("100")).quantize(Decimal("0.1")) if total_top_sales_amount > 0 else Decimal("0")

    top_waste_rows = sorted(
        [
            {"product": name, "amount": values["amount"], "qty": values["qty"]}
            for name, values in waste_by_product.items()
        ],
        key=lambda row: (-row["amount"], row["product"]),
    )[:10]
    total_top_waste_amount = sum((row["amount"] for row in top_waste_rows), Decimal("0"))
    for row in top_waste_rows:
        row["share_pct"] = ((row["amount"] / total_top_waste_amount) * Decimal("100")).quantize(Decimal("0.1")) if total_top_waste_amount > 0 else Decimal("0")

    category_rows = sorted(
        set(sales_by_category.keys()) | set(production_by_category.keys()),
        key=lambda label: (
            _POINT_DASHBOARD_CATEGORY_ORDER.index(label) if label in _POINT_DASHBOARD_CATEGORY_ORDER else 999,
            -(sales_by_category.get(label, Decimal("0")) + production_by_category.get(label, Decimal("0"))),
            label,
        ),
    )
    category_rows = [
        label
        for label in category_rows
        if sales_by_category.get(label, Decimal("0")) > 0 or production_by_category.get(label, Decimal("0")) > 0
    ][:10]
    unclassified_sales_qty = sales_by_category.get("Sin categoría", Decimal("0"))
    unclassified_production_qty = production_by_category.get("Sin categoría", Decimal("0"))
    classified_sales_qty = sum(value for key, value in sales_by_category.items() if key != "Sin categoría")
    classified_production_qty = sum(value for key, value in production_by_category.items() if key != "Sin categoría")
    category_balance_rows = []
    for label in category_rows:
        production_qty = production_by_category.get(label, Decimal("0"))
        sales_qty = sales_by_category.get(label, Decimal("0"))
        delta = production_qty - sales_qty
        if delta > 0:
            status = "Exceso"
            tone = "warning"
        elif delta < 0:
            status = "Faltante"
            tone = "danger"
        else:
            status = "Balanceado"
            tone = "success"
        category_balance_rows.append(
            {
                "label": label,
                "production_qty": production_qty,
                "sales_qty": sales_qty,
                "delta": delta,
                "status": status,
                "tone": tone,
            }
        )

    forecast_rows = []
    ranked_forecast_categories = sorted(
        next_week_categories.keys(),
        key=lambda label: (
            -(
                sum(next_week_categories[label]["sales_qty_values"], Decimal("0"))
                + sum(next_week_categories[label]["production_values"], Decimal("0"))
            ),
            label,
        ),
    )[:10]
    for label in ranked_forecast_categories:
        sales_amount_values: list[Decimal] = next_week_categories[label]["sales_amount_values"]  # type: ignore[assignment]
        sales_qty_values: list[Decimal] = next_week_categories[label]["sales_qty_values"]  # type: ignore[assignment]
        production_values: list[Decimal] = next_week_categories[label]["production_values"]  # type: ignore[assignment]
        avg_sales_amount = (sum(sales_amount_values, Decimal("0")) / Decimal(str(len(sales_amount_values)))) if sales_amount_values else Decimal("0")
        avg_sales_qty = (sum(sales_qty_values, Decimal("0")) / Decimal(str(len(sales_qty_values)))) if sales_qty_values else Decimal("0")
        avg_production_qty = (sum(production_values, Decimal("0")) / Decimal(str(len(production_values)))) if production_values else Decimal("0")
        trend = "up" if avg_sales_qty >= avg_production_qty else "down"
        forecast_rows.append(
            {
                "label": label,
                "forecast_sales_amount": avg_sales_amount.quantize(Decimal("0.01")),
                "forecast_sales_qty": avg_sales_qty.quantize(Decimal("0.001")),
                "forecast_production_qty": avg_production_qty.quantize(Decimal("0.001")),
                "trend": trend,
            }
        )

    weekly_rows = [weekly[week_start] for week_start in sorted(week_starts, reverse=True)]
    forecast_total_sales_amount = sum((row["sales_amount"] for row in weekly_rows), Decimal("0")) / Decimal(str(len(weekly_rows))) if weekly_rows else Decimal("0")
    forecast_total_units = sum((row["sales_units"] for row in weekly_rows), Decimal("0")) / Decimal(str(len(weekly_rows))) if weekly_rows else Decimal("0")
    forecast_total_production = sum((row["production_units"] for row in weekly_rows), Decimal("0")) / Decimal(str(len(weekly_rows))) if weekly_rows else Decimal("0")
    sales_category_total = classified_sales_qty + unclassified_sales_qty
    inflow_category_total = classified_production_qty + unclassified_production_qty
    sales_classification_pct = (
        (classified_sales_qty / sales_category_total * Decimal("100"))
        if sales_category_total > 0
        else Decimal("100")
    )
    inflow_classification_pct = (
        (classified_production_qty / inflow_category_total * Decimal("100"))
        if inflow_category_total > 0
        else Decimal("100")
    )
    inflow_gap_units = current["inflow_units"] - current["sales_units"]
    transfer_share_pct = (
        current["transfer_units"] / current["inflow_units"] * Decimal("100")
        if current["inflow_units"] > 0
        else Decimal("0")
    )
    waste_rate_pct = (
        current["waste_units"] / current["sales_units"] * Decimal("100")
        if current["sales_units"] > 0
        else Decimal("0")
    )

    coverage = current["coverage_ratio"]
    waste_trend = current["waste_cost"] - previous["waste_cost"]
    if coverage >= Decimal("1.00") and waste_trend <= 0:
        alert_status = "Verde"
        alert_tone = "success"
        alert_detail = "Cobertura operativa suficiente y merma estable o bajando."
    elif coverage >= Decimal("0.85"):
        alert_status = "Amarillo"
        alert_tone = "warning"
        alert_detail = "Operación relativamente controlada, pero conviene revisar presión comercial o merma."
    else:
        alert_status = "Rojo"
        alert_tone = "danger"
        alert_detail = "Entradas centrales por debajo de ventas semanales; conviene replanear producción."

    chart_bundle = {
        "weekly_labels": [row["week_label"] for row in reversed(weekly_rows)],
        "weekly_sales_amount": [float(row["sales_amount"]) for row in reversed(weekly_rows)],
        "weekly_sales_units": [float(row["sales_units"]) for row in reversed(weekly_rows)],
        "weekly_production_units": [float(row["production_units"]) for row in reversed(weekly_rows)],
        "weekly_waste_cost": [float(row["waste_cost"]) for row in reversed(weekly_rows)],
        "category_labels": [row["label"] for row in category_balance_rows],
        "category_production": [float(row["production_qty"]) for row in category_balance_rows],
        "category_sales": [float(row["sales_qty"]) for row in category_balance_rows],
        "waste_branch_labels": [label for label in waste_by_branch.keys()],
        "waste_branch_values": [float(values["cost"]) for values in waste_by_branch.values()],
        "forecast_labels": [row["label"] for row in forecast_rows],
        "forecast_current_production": [float(production_by_category.get(row["label"], Decimal("0"))) for row in forecast_rows],
        "forecast_next_production": [float(row["forecast_production_qty"]) for row in forecast_rows],
    }

    return {
        "week_start": current_week_start,
        "week_end": current_week_end,
        "week_label": _iso_week_label(current_week_start),
        "previous_week_label": _iso_week_label(week_starts[-2]),
        "coverage_ratio": coverage,
        "sales_amount": current["sales_amount"],
        "sales_units": current["sales_units"],
        "tickets": current["tickets"],
        "tickets_available": current["tickets"] > 0,
        "avg_ticket": current["avg_ticket"],
        "production_units": current["production_units"],
        "transfer_units": current["transfer_units"],
        "waste_cost": current["waste_cost"],
        "waste_units": current["waste_units"],
        "alert_status": alert_status,
        "alert_tone": alert_tone,
        "alert_detail": alert_detail,
        "week_comparison_rows": [
            {"label": "Ventas ($)", "current": current["sales_amount"], "previous": previous["sales_amount"], "delta_pct": _safe_delta_pct(current["sales_amount"], previous["sales_amount"])},
            {"label": "Unidades vendidas", "current": current["sales_units"], "previous": previous["sales_units"], "delta_pct": _safe_delta_pct(current["sales_units"], previous["sales_units"])},
            {"label": "Producción CEDIS", "current": current["production_units"], "previous": previous["production_units"], "delta_pct": _safe_delta_pct(current["production_units"], previous["production_units"])},
            {"label": "Transferencias a CEDIS", "current": current["transfer_units"], "previous": previous["transfer_units"], "delta_pct": _safe_delta_pct(current["transfer_units"], previous["transfer_units"])},
            {"label": "Ticket promedio", "current": current["avg_ticket"], "previous": previous["avg_ticket"], "delta_pct": _safe_delta_pct(current["avg_ticket"], previous["avg_ticket"])},
            {"label": "Merma ($)", "current": current["waste_cost"], "previous": previous["waste_cost"], "delta_pct": _safe_delta_pct(current["waste_cost"], previous["waste_cost"])},
            {"label": "Merma (pzas)", "current": current["waste_units"], "previous": previous["waste_units"], "delta_pct": _safe_delta_pct(current["waste_units"], previous["waste_units"])},
            {"label": "Cobertura operativa", "current": current["coverage_ratio"], "previous": previous["coverage_ratio"], "delta_pct": _safe_delta_pct(current["coverage_ratio"], previous["coverage_ratio"])},
        ],
        "weekly_rows": weekly_rows,
        "top_sales_rows": top_sales_rows,
        "top_waste_rows": top_waste_rows,
        "category_balance_rows": category_balance_rows,
        "waste_branch_rows": [
            {"label": label, "cost": values["cost"], "qty": values["qty"]}
            for label, values in sorted(waste_by_branch.items(), key=lambda item: (-item[1]["cost"], item[0]))
        ],
        "forecast_rows": forecast_rows,
        "forecast_total_sales_amount": forecast_total_sales_amount.quantize(Decimal("0.01")),
        "forecast_total_units": forecast_total_units.quantize(Decimal("0.001")),
        "forecast_total_production": forecast_total_production.quantize(Decimal("0.001")),
        "chart_bundle": chart_bundle,
        "executive_summary": {
            "inflow_units": current["inflow_units"],
            "inflow_gap_units": inflow_gap_units,
            "transfer_share_pct": transfer_share_pct.quantize(Decimal("0.1")),
            "waste_rate_pct": waste_rate_pct.quantize(Decimal("0.1")),
            "gap_tone": "success" if inflow_gap_units >= 0 else "danger",
            "gap_label": "Holgura central" if inflow_gap_units >= 0 else "Déficit central",
            "ticket_signal": "Confirmado" if current["tickets"] > 0 else "No disponible",
        },
        "category_mapping_quality": {
            "unclassified_sales_qty": unclassified_sales_qty,
            "unclassified_production_qty": unclassified_production_qty,
            "classified_sales_qty": classified_sales_qty,
            "classified_production_qty": classified_production_qty,
            "sales_classification_pct": sales_classification_pct.quantize(Decimal("0.1")),
            "inflow_classification_pct": inflow_classification_pct.quantize(Decimal("0.1")),
        },
    }


def _export_dg_operacion_dashboard_csv(payload: Dict[str, Any]) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="dg_operacion_cockpit.csv"'
    writer = csv.writer(response)

    writer.writerow(["COCKPIT DG OPERACION INTEGRADA"])
    writer.writerow(["Fecha operacion", payload["fecha_operacion"]])
    writer.writerow(["Fecha inicio planes", payload.get("dg_start_date") or ""])
    writer.writerow(["Fecha fin planes", payload.get("dg_end_date") or ""])
    writer.writerow(["Agrupacion planes", payload["plan_status_dashboard"].get("group_by_label") or "Fecha producción"])
    writer.writerow([])
    writer.writerow(["PLAN DE PRODUCCION"])
    writer.writerow(["Estatus", payload["plan_status_dashboard"]["status"]])
    writer.writerow(["Detalle", payload["plan_status_dashboard"]["detail"]])
    writer.writerow(["Abiertos", payload["plan_status_dashboard"]["abiertos"]])
    writer.writerow(["Cerrados", payload["plan_status_dashboard"]["cerrados"]])
    writer.writerow([])
    writer.writerow(
        [
            payload["plan_status_dashboard"].get("group_by_label") or "Fecha producción",
            "Total",
            "Borrador",
            "Consumo aplicado",
            "Cerrado",
            "Abiertos",
        ]
    )
    for row in payload["plan_status_dashboard"]["rows"]:
        writer.writerow([row["label"], row["total"], row["borrador"], row["consumo_aplicado"], row["cerrado"], row["abiertos"]])
    writer.writerow([])
    writer.writerow(["REABASTO CEDIS"])
    writer.writerow(["Semaforo", payload["reabasto_stage"]])
    writer.writerow(["Detalle", payload["reabasto_detail"]])
    writer.writerow(["Sucursales activas", payload["resumen_cierre"]["total"]])
    writer.writerow(["En tiempo", payload["resumen_cierre"]["en_tiempo"]])
    writer.writerow(["Tardias", payload["resumen_cierre"]["tardias"]])
    writer.writerow(["Pendientes", payload["resumen_cierre"]["pendientes"]])
    writer.writerow([])
    writer.writerow(["Sucursal", "Estatus", "Actualizado en"])
    for row in payload["resumen_cierre"]["detalle"]:
        writer.writerow([f"{row['sucursal'].codigo} · {row['sucursal'].nombre}", row["estado_label"], row["actualizado_en"] or ""])
    writer.writerow([])
    writer.writerow(["VENTAS HISTORICAS"])
    writer.writerow(["Estatus", payload["ventas_historicas_summary"]["status"]])
    writer.writerow(["Detalle", payload["ventas_historicas_summary"]["detail"]])
    writer.writerow(["Cobertura", payload["ventas_historicas_summary"]["date_label"]])
    writer.writerow(["Dias activos", payload["ventas_historicas_summary"]["active_days"]])
    writer.writerow(["Sucursales", payload["ventas_historicas_summary"]["branch_count"]])
    writer.writerow(["Recetas", payload["ventas_historicas_summary"]["recipe_count"]])
    writer.writerow(["Dias faltantes", payload["ventas_historicas_summary"]["missing_days"]])
    writer.writerow([])
    writer.writerow(["MERMAS POINT"])
    writer.writerow(["Fecha cierre", payload["point_waste_summary"]["closure_date"]])
    writer.writerow(["Estatus", payload["point_waste_summary"]["status"]])
    writer.writerow(["Detalle", payload["point_waste_summary"]["detail"]])
    writer.writerow(["Registros", payload["point_waste_summary"]["total_rows"]])
    writer.writerow(["Cantidad total", payload["point_waste_summary"]["total_qty"]])
    writer.writerow(["Sucursales con merma", payload["point_waste_summary"]["branch_count"]])
    writer.writerow(["Responsables", payload["point_waste_summary"]["responsible_count"]])
    writer.writerow([])
    writer.writerow(["FLUJO CENTRAL POINT"])
    writer.writerow(["Fecha cierre", payload["point_central_flow_summary"]["closure_date"]])
    writer.writerow(["Estatus", payload["point_central_flow_summary"]["status"]])
    writer.writerow(["Detalle", payload["point_central_flow_summary"]["detail"]])
    writer.writerow(["Produccion central lineas", payload["point_central_flow_summary"]["production_rows"]])
    writer.writerow(["Produccion central qty", payload["point_central_flow_summary"]["production_total_qty"]])
    writer.writerow(["Produccion central items", payload["point_central_flow_summary"]["production_item_count"]])
    writer.writerow(["Transferencias recibidas lineas", payload["point_central_flow_summary"]["transfer_rows"]])
    writer.writerow(["Transferencias recibidas qty", payload["point_central_flow_summary"]["transfer_total_qty"]])
    writer.writerow(["Transferencias recibidas items", payload["point_central_flow_summary"]["transfer_item_count"]])
    writer.writerow([])
    writer.writerow(["CUADRE OPERATIVO POINT"])
    writer.writerow(["Fecha cierre", payload["point_closure_summary"]["closure_date"]])
    writer.writerow(["Estatus", payload["point_closure_summary"]["status"]])
    writer.writerow(["Detalle", payload["point_closure_summary"]["detail"]])
    writer.writerow(["Sucursales evaluadas", payload["point_closure_summary"]["branch_count"]])
    writer.writerow(["Cuadra", payload["point_closure_summary"]["cuadra"]])
    writer.writerow(["Desviacion menor", payload["point_closure_summary"]["minor"]])
    writer.writerow(["Desviacion critica", payload["point_closure_summary"]["critical"]])
    writer.writerow(["Sin cierre", payload["point_closure_summary"]["without_closure"]])
    writer.writerow([])
    writer.writerow(["Sucursal", "Inicial", "Produccion", "Transferencias", "Ventas", "Tickets", "Ticket promedio", "Mermas", "Esperado", "Cierre", "Variacion", "Estatus"])
    for row in payload["point_closure_summary"]["rows"]:
        writer.writerow(
            [
                row["branch_label"],
                row["opening_units"],
                row["produced_units"],
                row["transfer_units"],
                row["sold_units"],
                row["sold_tickets"],
                row["avg_ticket"],
                row["waste_units"],
                row["expected_units"],
                row["closing_units"],
                row["variance_units"],
                row["status"],
            ]
        )
    return response


def _export_dg_operacion_dashboard_xlsx(payload: Dict[str, Any]) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "DG Operacion"
    ws.append(["COCKPIT DG OPERACION INTEGRADA", ""])
    ws.append(["Fecha operacion", str(payload["fecha_operacion"])])
    ws.append(["Fecha inicio planes", payload.get("dg_start_date") or ""])
    ws.append(["Fecha fin planes", payload.get("dg_end_date") or ""])
    ws.append(["Agrupacion planes", payload["plan_status_dashboard"].get("group_by_label") or "Fecha producción"])
    ws.append([])
    ws.append(["PLAN DE PRODUCCION", ""])
    ws.append(["Estatus", payload["plan_status_dashboard"]["status"]])
    ws.append(["Detalle", payload["plan_status_dashboard"]["detail"]])
    ws.append(["Abiertos", int(payload["plan_status_dashboard"]["abiertos"])])
    ws.append(["Cerrados", int(payload["plan_status_dashboard"]["cerrados"])])
    ws.append([])
    ws.append(
        [
            payload["plan_status_dashboard"].get("group_by_label") or "Fecha producción",
            "Total",
            "Borrador",
            "Consumo aplicado",
            "Cerrado",
            "Abiertos",
        ]
    )
    for row in payload["plan_status_dashboard"]["rows"]:
        ws.append([str(row["label"]), int(row["total"]), int(row["borrador"]), int(row["consumo_aplicado"]), int(row["cerrado"]), int(row["abiertos"])])
    ws.append([])
    ws.append(["REABASTO CEDIS", ""])
    ws.append(["Semaforo", payload["reabasto_stage"]])
    ws.append(["Detalle", payload["reabasto_detail"]])
    ws.append(["Sucursales activas", int(payload["resumen_cierre"]["total"])])
    ws.append(["En tiempo", int(payload["resumen_cierre"]["en_tiempo"])])
    ws.append(["Tardias", int(payload["resumen_cierre"]["tardias"])])
    ws.append(["Pendientes", int(payload["resumen_cierre"]["pendientes"])])
    ws.append([])
    ws.append(["Sucursal", "Estatus", "Actualizado en"])
    for row in payload["resumen_cierre"]["detalle"]:
        ws.append([f"{row['sucursal'].codigo} · {row['sucursal'].nombre}", row["estado_label"], str(row["actualizado_en"] or "")])
    ws.append([])
    ws.append(["VENTAS HISTORICAS", ""])
    ws.append(["Estatus", payload["ventas_historicas_summary"]["status"]])
    ws.append(["Detalle", payload["ventas_historicas_summary"]["detail"]])
    ws.append(["Cobertura", payload["ventas_historicas_summary"]["date_label"]])
    ws.append(["Dias activos", int(payload["ventas_historicas_summary"]["active_days"])])
    ws.append(["Sucursales", int(payload["ventas_historicas_summary"]["branch_count"])])
    ws.append(["Recetas", int(payload["ventas_historicas_summary"]["recipe_count"])])
    ws.append(["Dias faltantes", int(payload["ventas_historicas_summary"]["missing_days"])])
    ws.append([])
    ws.append(["MERMAS POINT", ""])
    ws.append(["Fecha cierre", str(payload["point_waste_summary"]["closure_date"])])
    ws.append(["Estatus", payload["point_waste_summary"]["status"]])
    ws.append(["Detalle", payload["point_waste_summary"]["detail"]])
    ws.append(["Registros", int(payload["point_waste_summary"]["total_rows"])])
    ws.append(["Cantidad total", float(payload["point_waste_summary"]["total_qty"])])
    ws.append(["Sucursales con merma", int(payload["point_waste_summary"]["branch_count"])])
    ws.append(["Responsables", int(payload["point_waste_summary"]["responsible_count"])])
    ws.append([])
    ws.append(["FLUJO CENTRAL POINT", ""])
    ws.append(["Fecha cierre", str(payload["point_central_flow_summary"]["closure_date"])])
    ws.append(["Estatus", payload["point_central_flow_summary"]["status"]])
    ws.append(["Detalle", payload["point_central_flow_summary"]["detail"]])
    ws.append(["Produccion central lineas", int(payload["point_central_flow_summary"]["production_rows"])])
    ws.append(["Produccion central qty", float(payload["point_central_flow_summary"]["production_total_qty"])])
    ws.append(["Produccion central items", int(payload["point_central_flow_summary"]["production_item_count"])])
    ws.append(["Transferencias recibidas lineas", int(payload["point_central_flow_summary"]["transfer_rows"])])
    ws.append(["Transferencias recibidas qty", float(payload["point_central_flow_summary"]["transfer_total_qty"])])
    ws.append(["Transferencias recibidas items", int(payload["point_central_flow_summary"]["transfer_item_count"])])
    ws.append([])
    ws.append(["CUADRE OPERATIVO POINT", ""])
    ws.append(["Fecha cierre", str(payload["point_closure_summary"]["closure_date"])])
    ws.append(["Estatus", payload["point_closure_summary"]["status"]])
    ws.append(["Detalle", payload["point_closure_summary"]["detail"]])
    ws.append(["Sucursales evaluadas", int(payload["point_closure_summary"]["branch_count"])])
    ws.append(["Cuadra", int(payload["point_closure_summary"]["cuadra"])])
    ws.append(["Desviacion menor", int(payload["point_closure_summary"]["minor"])])
    ws.append(["Desviacion critica", int(payload["point_closure_summary"]["critical"])])
    ws.append(["Sin cierre", int(payload["point_closure_summary"]["without_closure"])])
    ws.append([])
    ws.append(["Sucursal", "Inicial", "Produccion", "Transferencias", "Ventas", "Tickets", "Ticket promedio", "Mermas", "Esperado", "Cierre", "Variacion", "Estatus"])
    for row in payload["point_closure_summary"]["rows"]:
        ws.append(
            [
                row["branch_label"],
                float(row["opening_units"]),
                float(row["produced_units"]),
                float(row["transfer_units"]),
                float(row["sold_units"]),
                int(row["sold_tickets"]),
                float(row["avg_ticket"]),
                float(row["waste_units"]),
                float(row["expected_units"]),
                float(row["closing_units"]),
                float(row["variance_units"]),
                row["status"],
            ]
        )
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="dg_operacion_cockpit.xlsx"'
    return response


@login_required
@permission_required("recetas.view_planproduccion", raise_exception=True)
def dg_operacion_dashboard(request: HttpRequest) -> HttpResponse:
    dg_filters = _plan_status_dashboard_filters(request)
    payload = _build_dg_operacion_dashboard_payload(
        start_date=dg_filters["start_date"],
        end_date=dg_filters["end_date"],
        group_by=dg_filters["group_by"],
        fecha_operacion=_parse_date_safe(request.GET.get("fecha_operacion")),
    )

    return render(
        request,
        "recetas/dg_operacion_dashboard.html",
        payload,
    )


@login_required
@permission_required("recetas.view_planproduccion", raise_exception=True)
def dg_operacion_dashboard_export(request: HttpRequest) -> HttpResponse:
    dg_filters = _plan_status_dashboard_filters(request)
    payload = _build_dg_operacion_dashboard_payload(
        start_date=dg_filters["start_date"],
        end_date=dg_filters["end_date"],
        group_by=dg_filters["group_by"],
        fecha_operacion=_parse_date_safe(request.GET.get("fecha_operacion")),
    )
    export_format = (request.GET.get("format") or "csv").strip().lower()
    if export_format == "xlsx":
        return _export_dg_operacion_dashboard_xlsx(payload)
    return _export_dg_operacion_dashboard_csv(payload)


@login_required
@permission_required("recetas.view_planproduccion", raise_exception=True)
def produccion_cedis_weekly_dashboard(request: HttpRequest) -> HttpResponse:
    base_date = _parse_date_safe(request.GET.get("week_of")) or _latest_point_operational_cutoff_date()
    payload = _build_produccion_cedis_weekly_dashboard(base_date)
    return render(
        request,
        "recetas/produccion_cedis_weekly_dashboard.html",
        payload,
    )


@login_required
@permission_required("recetas.view_planproduccion", raise_exception=True)
def plan_produccion_solicitud_print(request: HttpRequest, plan_id: int) -> HttpResponse:
    plan = get_object_or_404(PlanProduccion, pk=plan_id)
    explosion = _plan_explosion(plan)
    materias_primas = [row for row in explosion["insumos"] if row["origen"] != "Interno"]
    internos = [row for row in explosion["insumos"] if row["origen"] == "Interno"]
    folio = f"SOL-{plan.fecha_produccion.strftime('%Y%m%d')}-{plan.id:04d}"
    return render(
        request,
        "recetas/solicitud_insumos_print.html",
        {
            "plan": plan,
            "explosion": explosion,
            "materias_primas": materias_primas,
            "internos": internos,
            "folio": folio,
            "titulo_documento": "Orden de Solicitud de Insumos",
            "subtitulo_documento": "Documento interno de abastecimiento",
            "mostrar_internos": True,
        },
    )


@login_required
@permission_required("recetas.view_planproduccion", raise_exception=True)
def plan_produccion_solicitud_compras_print(request: HttpRequest, plan_id: int) -> HttpResponse:
    plan = get_object_or_404(PlanProduccion, pk=plan_id)
    explosion = _plan_explosion(plan)
    materias_primas = [row for row in explosion["insumos"] if row["origen"] != "Interno"]
    folio = f"SOL-COMP-{plan.fecha_produccion.strftime('%Y%m%d')}-{plan.id:04d}"
    return render(
        request,
        "recetas/solicitud_insumos_print.html",
        {
            "plan": plan,
            "explosion": explosion,
            "materias_primas": materias_primas,
            "internos": [],
            "folio": folio,
            "titulo_documento": "Orden de Solicitud de Materia Prima",
            "subtitulo_documento": "Documento para área de Compras",
            "mostrar_internos": False,
        },
    )


@login_required
@permission_required("recetas.add_planproduccion", raise_exception=True)
@require_POST
def plan_produccion_create(request: HttpRequest) -> HttpResponse:
    nombre = (request.POST.get("nombre") or "").strip()
    fecha_str = (request.POST.get("fecha_produccion") or "").strip()
    notas = (request.POST.get("notas") or "").strip()
    if not nombre:
        messages.error(request, "El nombre del plan es obligatorio.")
        return redirect("recetas:plan_produccion")

    fecha = timezone.localdate()
    if fecha_str:
        try:
            fecha = date.fromisoformat(fecha_str)
        except Exception:
            messages.warning(request, "Fecha inválida. Se usó la fecha de hoy.")

    plan = PlanProduccion.objects.create(
        nombre=nombre[:140],
        fecha_produccion=fecha,
        notas=notas,
        creado_por=request.user if request.user.is_authenticated else None,
    )
    messages.success(request, "Plan de producción creado.")
    return redirect(f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}")


@login_required
@permission_required("recetas.change_planproduccion", raise_exception=True)
@require_POST
def plan_produccion_item_create(request: HttpRequest, plan_id: int) -> HttpResponse:
    plan = get_object_or_404(PlanProduccion, pk=plan_id)
    receta_id = request.POST.get("receta_id")
    cantidad_raw = (request.POST.get("cantidad") or "1").strip()
    notas = (request.POST.get("notas") or "").strip()[:160]

    receta = Receta.objects.filter(pk=receta_id).first()
    if not receta:
        messages.error(request, "Selecciona una receta válida.")
        return redirect(f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}")

    try:
        cantidad = Decimal(cantidad_raw)
    except Exception:
        cantidad = Decimal("0")
    if cantidad <= 0:
        messages.error(request, "La cantidad debe ser mayor que cero.")
        return redirect(f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}")

    PlanProduccionItem.objects.create(
        plan=plan,
        receta=receta,
        cantidad=cantidad,
        notas=notas,
    )
    messages.success(request, "Producto agregado al plan.")
    return redirect(f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}")


@login_required
@permission_required("recetas.change_planproduccion", raise_exception=True)
@require_POST
def plan_produccion_item_delete(request: HttpRequest, plan_id: int, item_id: int) -> HttpResponse:
    plan = get_object_or_404(PlanProduccion, pk=plan_id)
    item = get_object_or_404(PlanProduccionItem, pk=item_id, plan=plan)
    item.delete()
    messages.success(request, "Renglón eliminado del plan.")
    return redirect(f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}")


@login_required
@permission_required("recetas.delete_planproduccion", raise_exception=True)
@require_POST
def plan_produccion_delete(request: HttpRequest, plan_id: int) -> HttpResponse:
    plan = get_object_or_404(PlanProduccion, pk=plan_id)
    plan.delete()
    messages.success(request, "Plan eliminado.")
    return redirect("recetas:plan_produccion")


@login_required
@require_POST
def plan_produccion_generar_solicitudes(request: HttpRequest, plan_id: int) -> HttpResponse:
    plan = get_object_or_404(PlanProduccion, pk=plan_id)
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para generar solicitudes de compra.")

    auto_create_oc = bool(request.POST.get("auto_create_oc"))
    replace_prev_raw = (request.POST.get("replace_prev") or "1").strip().lower()
    replace_prev = replace_prev_raw not in {"0", "false", "off", "no"}
    stats = _generar_solicitudes_compra_desde_plan(
        plan=plan,
        user=request.user,
        replace_prev=replace_prev,
        auto_create_oc=auto_create_oc,
    )

    if stats["creadas"] == 0 and stats["actualizadas"] == 0:
        messages.warning(
            request,
            "No se generaron solicitudes: el plan no tiene materia prima con cantidad válida.",
        )
    else:
        mode_label = "reemplazo" if replace_prev else "acumulado"
        msg = (
            f"Solicitudes generadas: {stats['creadas']}. "
            f"Solicitudes actualizadas: {stats['actualizadas']}. "
            f"Modo: {mode_label}. "
            f"Borradores reemplazados del plan: {stats['deleted_prev']}."
        )
        if auto_create_oc:
            msg += (
                f" OC borrador creadas (agrupadas por proveedor): {stats['oc_creadas']}. "
                f"OC borrador actualizadas: {stats['oc_actualizadas']}. "
                f"OCs borrador previas reemplazadas: {stats['oc_prev_deleted']}."
            )
            if stats["sin_proveedor"]:
                msg += (
                    f" Insumos sin proveedor principal: {stats['sin_proveedor']} "
                    "(no entraron a OC automática)."
                )
        messages.success(request, msg)
    next_view = (request.POST.get("next_view") or "plan").strip().lower()
    compras_query = urlencode(
        {
            "source": "plan",
            "plan_id": str(plan.id),
            "reabasto": "all",
        }
    )
    if next_view == "compras":
        return redirect(f"{reverse('compras:solicitudes')}?{compras_query}")
    if next_view == "compras_print":
        return redirect(f"{reverse('compras:solicitudes_print')}?{compras_query}")
    return redirect(f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}")


@login_required
@permission_required("recetas.change_planproduccion", raise_exception=True)
@require_POST
def plan_produccion_aplicar_consumo(request: HttpRequest, plan_id: int) -> HttpResponse:
    plan = get_object_or_404(PlanProduccion, pk=plan_id)
    try:
        stats = _apply_plan_consumption(plan, request.user)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect(f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}")

    if (
        stats["insumos_created"] == 0
        and stats["productos_created"] == 0
        and (stats["insumos_skipped"] > 0 or stats["productos_skipped"] > 0)
    ):
        messages.warning(
            request,
            "El consumo del plan ya estaba aplicado. No se descontó inventario nuevamente.",
        )
    else:
        messages.success(
            request,
            (
                "Consumo aplicado al inventario del plan. "
                f"Insumos descontados: {stats['insumos_created']}. "
                f"Productos padre descontados: {stats['productos_created']}. "
                f"Saltados por idempotencia: {stats['insumos_skipped'] + stats['productos_skipped']}."
            ),
        )
    return redirect(f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}")


@login_required
@permission_required("recetas.change_planproduccion", raise_exception=True)
@require_POST
def plan_produccion_cerrar(request: HttpRequest, plan_id: int) -> HttpResponse:
    plan = get_object_or_404(PlanProduccion, pk=plan_id)
    if plan.estado == PlanProduccion.ESTADO_CERRADO:
        messages.warning(request, "El plan ya estaba cerrado.")
        return redirect(f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}")

    if not plan.consumo_aplicado:
        messages.error(request, "No puedes cerrar el plan mientras el consumo real de inventario siga pendiente.")
        return redirect(f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}")

    document_control = _plan_document_control(plan)
    blocked_total = int((document_control or {}).get("blocked_total") or 0)
    closure_pending = int(((document_control or {}).get("closure_summary") or {}).get("pending_count") or 0)
    if blocked_total > 0 or closure_pending > 0:
        messages.error(
            request,
            "No puedes cerrar el plan mientras existan bloqueos o criterios documentales pendientes.",
        )
        return redirect(f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}")

    with transaction.atomic():
        _mark_plan_closed(plan, request.user)
    messages.success(request, "Plan cerrado operativamente. Quedó listo para reporteo y auditoría.")
    return redirect(f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}")


def _generar_solicitudes_compra_desde_plan(
    plan: PlanProduccion,
    user,
    replace_prev: bool = True,
    auto_create_oc: bool = False,
    area_tag: str | None = None,
    referencia_plan: str | None = None,
) -> Dict[str, Any]:
    explosion = _plan_explosion(plan)
    materias_primas = [
        row
        for row in explosion["insumos"]
        if row["origen"] != "Interno" and Decimal(str(row["cantidad"] or 0)) > 0
    ]
    area_tag = area_tag or f"PLAN_PRODUCCION:{plan.id}"
    referencia_plan = referencia_plan or f"PLAN_PRODUCCION:{plan.id}"

    deleted_prev = 0
    if replace_prev:
        borradores_previos = SolicitudCompra.objects.filter(
            area=area_tag,
            estatus=SolicitudCompra.STATUS_BORRADOR,
        )
        deleted_prev = borradores_previos.count()
        if deleted_prev:
            borradores_previos.delete()

    creadas = 0
    actualizadas = 0
    sin_proveedor = 0
    oc_por_proveedor: dict[int, dict[str, Any]] = {}
    for row in materias_primas:
        insumo = Insumo.objects.filter(pk=row["insumo_id"]).first()
        if not insumo:
            continue
        proveedor = insumo.proveedor_principal
        qty = Decimal(str(row["cantidad"]))

        solicitud = None
        if not replace_prev:
            solicitud = (
                SolicitudCompra.objects.filter(
                    area=area_tag,
                    estatus=SolicitudCompra.STATUS_BORRADOR,
                    insumo=insumo,
                    fecha_requerida=plan.fecha_produccion,
                )
                .order_by("-creado_en")
                .first()
            )

        if solicitud:
            solicitud.cantidad = Decimal(str(solicitud.cantidad or 0)) + qty
            if not solicitud.proveedor_sugerido_id and proveedor:
                solicitud.proveedor_sugerido = proveedor
                solicitud.save(update_fields=["cantidad", "proveedor_sugerido"])
            else:
                solicitud.save(update_fields=["cantidad"])
            log_event(
                user,
                "UPDATE",
                "compras.SolicitudCompra",
                solicitud.id,
                {
                    "folio": solicitud.folio,
                    "source_plan_id": plan.id,
                    "source_plan_nombre": plan.nombre,
                    "mode": "accumulate",
                },
            )
            actualizadas += 1
        else:
            solicitud = SolicitudCompra.objects.create(
                area=area_tag,
                solicitante=user.username,
                insumo=insumo,
                proveedor_sugerido=proveedor,
                cantidad=qty,
                fecha_requerida=plan.fecha_produccion,
                estatus=SolicitudCompra.STATUS_BORRADOR,
            )
            log_event(
                user,
                "CREATE",
                "compras.SolicitudCompra",
                solicitud.id,
                {
                    "folio": solicitud.folio,
                    "source_plan_id": plan.id,
                    "source_plan_nombre": plan.nombre,
                },
            )
            creadas += 1

        if auto_create_oc:
            if not proveedor:
                sin_proveedor += 1
                continue
            bucket = oc_por_proveedor.setdefault(
                proveedor.id,
                {
                    "proveedor_id": proveedor.id,
                    "proveedor_nombre": proveedor.nombre,
                    "monto_estimado": Decimal("0"),
                },
            )
            bucket["monto_estimado"] += Decimal(str(row.get("costo_total") or 0))

    oc_prev_deleted = 0
    oc_creadas = 0
    oc_actualizadas = 0
    if auto_create_oc:
        if replace_prev:
            ocs_previas = OrdenCompra.objects.filter(
                referencia=referencia_plan,
                estatus=OrdenCompra.STATUS_BORRADOR,
                solicitud__isnull=True,
            )
            oc_prev_deleted = ocs_previas.count()
            if oc_prev_deleted:
                ocs_previas.delete()

        for data in oc_por_proveedor.values():
            orden = None
            if not replace_prev:
                orden = (
                    OrdenCompra.objects.filter(
                        referencia=referencia_plan,
                        estatus=OrdenCompra.STATUS_BORRADOR,
                        solicitud__isnull=True,
                        proveedor_id=data["proveedor_id"],
                    )
                    .order_by("-creado_en")
                    .first()
                )

            if orden:
                orden.monto_estimado = Decimal(str(orden.monto_estimado or 0)) + data["monto_estimado"]
                orden.fecha_entrega_estimada = plan.fecha_produccion
                orden.save(update_fields=["monto_estimado", "fecha_entrega_estimada"])
                log_event(
                    user,
                    "UPDATE",
                    "compras.OrdenCompra",
                    orden.id,
                    {
                        "folio": orden.folio,
                        "estatus": orden.estatus,
                        "source_plan_id": plan.id,
                        "source_plan_nombre": plan.nombre,
                        "proveedor": data["proveedor_nombre"],
                        "mode": "accumulate",
                    },
                )
                oc_actualizadas += 1
            else:
                orden = OrdenCompra.objects.create(
                    solicitud=None,
                    referencia=referencia_plan,
                    proveedor_id=data["proveedor_id"],
                    fecha_emision=timezone.localdate(),
                    fecha_entrega_estimada=plan.fecha_produccion,
                    monto_estimado=data["monto_estimado"],
                    estatus=OrdenCompra.STATUS_BORRADOR,
                )
                log_event(
                    user,
                    "CREATE",
                    "compras.OrdenCompra",
                    orden.id,
                    {
                        "folio": orden.folio,
                        "estatus": orden.estatus,
                        "source_plan_id": plan.id,
                        "source_plan_nombre": plan.nombre,
                        "proveedor": data["proveedor_nombre"],
                    },
                )
                oc_creadas += 1

    return {
        "creadas": creadas,
        "actualizadas": actualizadas,
        "deleted_prev": deleted_prev,
        "sin_proveedor": sin_proveedor,
        "oc_prev_deleted": oc_prev_deleted,
        "oc_creadas": oc_creadas,
        "oc_actualizadas": oc_actualizadas,
    }


def _reabasto_redirect(
    fecha_operacion: date,
    sucursal_id: int | None = None,
    capture_only: bool = False,
) -> str:
    params = {"fecha": fecha_operacion.isoformat()}
    if sucursal_id:
        params["sucursal_id"] = str(sucursal_id)
    route_name = "recetas:reabasto_cedis_captura" if capture_only else "recetas:reabasto_cedis"
    return f"{reverse(route_name)}?{urlencode(params)}"


def _quantize_qty(value: Decimal) -> Decimal:
    return max(Decimal("0"), Decimal(str(value or 0))).quantize(Decimal("0.001"))


def _aplicar_lote_multiplo(value: Decimal, policy: PoliticaStockSucursalProducto | None) -> Decimal:
    qty = _quantize_qty(value)
    if qty <= 0 or not policy:
        return qty

    lote_minimo = _quantize_qty(policy.lote_minimo)
    if lote_minimo > 0 and qty < lote_minimo:
        qty = lote_minimo

    multiplo = _quantize_qty(policy.multiplo_empaque)
    if multiplo > 0:
        factor = (qty / multiplo).quantize(Decimal("1"), rounding=ROUND_CEILING)
        qty = _quantize_qty(factor * multiplo)
    return qty


def _calcular_sugerido_reabasto(
    policy: PoliticaStockSucursalProducto | None,
    stock_reportado: Decimal,
    en_transito: Decimal,
    consumo_proyectado: Decimal,
) -> Decimal:
    if not policy:
        return Decimal("0")

    objetivo = _to_decimal_safe(policy.stock_objetivo)
    seguridad = _to_decimal_safe(policy.stock_seguridad)
    sugerido_base = objetivo + seguridad + consumo_proyectado - stock_reportado - en_transito
    return _aplicar_lote_multiplo(_quantize_qty(sugerido_base), policy)


def _map_reabasto_header(header: str) -> str:
    key = normalizar_nombre(header or "").replace("_", " ")
    if key in {"receta", "producto", "producto final", "nombre producto", "nombre receta", "item"}:
        return "receta"
    if key in {"codigo", "codigo point", "sku"}:
        return "codigo_point"
    if key in {
        "cantidad",
        "solicitado",
        "cantidad solicitada",
        "cantidad pedido",
        "pedido",
        "solicitud final",
        "solicitud_final",
        "pedido final",
    }:
        return "solicitado"
    if key in {
        "stock",
        "stock reportado",
        "existencia",
        "inventario",
        "existencia cierre",
        "existencia al cierre",
        "stock cierre",
        "stock final",
        "stock final cierre",
        "stock_final_cierre",
    }:
        return "stock_reportado"
    if key in {"en transito", "transito"}:
        return "en_transito"
    if key in {"consumo proyectado", "consumo", "proyeccion consumo"}:
        return "consumo_proyectado"
    if key in {"justificacion", "motivo"}:
        return "justificacion"
    if key in {"observaciones", "nota", "notas"}:
        return "observaciones"
    if key in {"solicitud sugerida", "cantidad sugerida"}:
        return "solicitado"
    if key in {"solicitud requerida", "cantidad requerida", "solicitud_requerida"}:
        return "solicitado"
    return key


def _load_reabasto_rows(uploaded) -> list[dict]:
    filename = (uploaded.name or "").lower()
    rows: list[dict] = []
    if filename.endswith(".csv"):
        uploaded.seek(0)
        content = uploaded.read().decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(content.splitlines())
        for raw in reader:
            parsed = {}
            for key, value in (raw or {}).items():
                if not key:
                    continue
                parsed[_map_reabasto_header(str(key))] = value
            rows.append(parsed)
        return rows

    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        uploaded.seek(0)
        wb = load_workbook(uploaded, read_only=True, data_only=True)
        ws = wb.active
        values = list(ws.values)
        if not values:
            return []
        headers = [_map_reabasto_header(str(h or "")) for h in values[0]]
        for raw in values[1:]:
            parsed = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                parsed[header] = raw[idx] if idx < len(raw) else None
            rows.append(parsed)
        return rows

    raise ValueError("Formato no soportado. Usa CSV o XLSX.")


def _resolve_receta_reabasto(receta_txt: str, codigo_point: str) -> Receta | None:
    code = (codigo_point or "").strip()
    if code:
        receta = Receta.objects.filter(codigo_point__iexact=code).first()
        if receta:
            return receta
    name = normalizar_nombre(receta_txt or "")
    if not name:
        return None
    return Receta.objects.filter(nombre_normalizado=name).order_by("id").first()


def _consolidado_reabasto_por_fecha(fecha_operacion: date) -> list[dict]:
    totals = (
        SolicitudReabastoCedisLinea.objects.filter(solicitud__fecha_operacion=fecha_operacion)
        .exclude(solicitud__estado=SolicitudReabastoCedis.ESTADO_CANCELADA)
        .values("receta_id", "receta__nombre")
        .annotate(
            total_solicitado=Sum("solicitado"),
            total_sugerido=Sum("sugerido"),
            sucursales=Count("solicitud__sucursal", distinct=True),
        )
        .order_by("receta__nombre")
    )
    inv_map = {
        x.receta_id: x
        for x in InventarioCedisProducto.objects.filter(receta_id__in=[row["receta_id"] for row in totals]).select_related("receta")
    }
    rows: list[dict] = []
    for row in totals:
        rid = row["receta_id"]
        inv = inv_map.get(rid)
        disponible = inv.disponible if inv else Decimal("0")
        solicitado = _to_decimal_safe(row.get("total_solicitado"))
        faltante = max(Decimal("0"), solicitado - disponible)
        rows.append(
            {
                "receta_id": rid,
                "receta": row.get("receta__nombre") or "-",
                "sucursales": int(row.get("sucursales") or 0),
                "total_sugerido": _quantize_qty(_to_decimal_safe(row.get("total_sugerido"))),
                "total_solicitado": _quantize_qty(solicitado),
                "cedis_disponible": _quantize_qty(disponible),
                "cedis_faltante_producir": _quantize_qty(faltante),
            }
        )
    return rows


def _resumen_cierre_sucursales_reabasto(fecha_operacion: date, sucursales: list[Sucursal]) -> Dict[str, Any]:
    fecha_corte = fecha_operacion - timedelta(days=1)
    solicitudes = (
        SolicitudReabastoCedis.objects.filter(fecha_operacion=fecha_operacion, sucursal__in=sucursales)
        .select_related("sucursal")
        .order_by("sucursal__codigo")
    )
    by_sucursal = {s.sucursal_id: s for s in solicitudes}

    detalle: list[dict] = []
    enviadas_en_tiempo = 0
    enviadas_tardias = 0
    borrador = 0
    pendientes = 0
    for suc in sucursales:
        sol = by_sucursal.get(suc.id)
        if not sol:
            pendientes += 1
            detalle.append(
                {
                    "sucursal": suc,
                    "estado": "PENDIENTE",
                    "estado_label": "Por validar (sin captura)",
                    "actualizado_en": None,
                    "semaforo": "rojo",
                }
            )
            continue

        actualizado_local = timezone.localtime(sol.actualizado_en) if sol.actualizado_en else None
        actualizado_date = actualizado_local.date() if actualizado_local else None
        estado = (sol.estado or "").upper()

        if estado in {SolicitudReabastoCedis.ESTADO_ENVIADA, SolicitudReabastoCedis.ESTADO_ATENDIDA}:
            if actualizado_date and actualizado_date <= fecha_corte:
                enviadas_en_tiempo += 1
                detalle.append(
                    {
                        "sucursal": suc,
                        "estado": estado,
                        "estado_label": f"{sol.get_estado_display()} (en tiempo)",
                        "actualizado_en": actualizado_local,
                        "semaforo": "verde",
                    }
                )
            else:
                enviadas_tardias += 1
                detalle.append(
                    {
                        "sucursal": suc,
                        "estado": estado,
                        "estado_label": f"{sol.get_estado_display()} (tardía)",
                        "actualizado_en": actualizado_local,
                        "semaforo": "amarillo",
                    }
                )
            continue

        if estado == SolicitudReabastoCedis.ESTADO_BORRADOR:
            borrador += 1
            detalle.append(
                {
                    "sucursal": suc,
                    "estado": estado,
                    "estado_label": "Borrador (no enviada)",
                    "actualizado_en": actualizado_local,
                    "semaforo": "amarillo",
                }
            )
        else:
            pendientes += 1
            detalle.append(
                {
                    "sucursal": suc,
                    "estado": estado or "PENDIENTE",
                    "estado_label": "Por validar (sin envío válido)",
                    "actualizado_en": actualizado_local,
                    "semaforo": "rojo",
                }
            )

    total = len(sucursales)
    listo_8am = total > 0 and enviadas_en_tiempo == total
    pendientes_codigos = [row["sucursal"].codigo for row in detalle if row["semaforo"] in {"rojo", "amarillo"}]
    return {
        "fecha_corte": fecha_corte,
        "total": total,
        "en_tiempo": enviadas_en_tiempo,
        "tardias": enviadas_tardias,
        "borrador": borrador,
        "pendientes": pendientes,
        "listo_8am": listo_8am,
        "detalle": detalle,
        "pendientes_codigos": pendientes_codigos,
    }


def _user_sucursal_scope(user) -> Sucursal | None:
    profile = getattr(user, "userprofile", None)
    if profile and getattr(profile, "sucursal_id", None):
        return profile.sucursal
    return None


def _assert_can_capture_for_sucursal(user, sucursal: Sucursal) -> None:
    """
    En flujo sucursal->CEDIS, un usuario no gestor solo puede operar su sucursal asignada.
    Si no tiene sucursal asignada, se bloquea para evitar capturas cruzadas.
    """
    if can_manage_compras(user):
        return
    user_sucursal = _user_sucursal_scope(user)
    if not user_sucursal:
        raise PermissionDenied("Tu usuario no tiene sucursal asignada para captura de reabasto.")
    if user_sucursal.id != sucursal.id:
        raise PermissionDenied("Solo puedes capturar reabasto para tu sucursal asignada.")


def _build_reabasto_rows(fecha_operacion: date, sucursal: Sucursal | None) -> dict[str, Any]:
    solicitud = None
    lineas = []
    politicas = []
    rows_detalle: list[dict] = []
    rows_cierre: list[dict] = []
    recetas_map: dict[int, Receta] = {}

    if sucursal:
        solicitud = (
            SolicitudReabastoCedis.objects.filter(
                fecha_operacion=fecha_operacion,
                sucursal=sucursal,
            )
            .select_related("sucursal", "creado_por")
            .first()
        )
        politicas = list(
            PoliticaStockSucursalProducto.objects.filter(sucursal=sucursal, activa=True)
            .select_related("receta")
            .order_by("receta__nombre")
        )
        if solicitud:
            lineas = list(
                solicitud.lineas.select_related("receta", "receta__rendimiento_unidad").order_by("receta__nombre")
            )

        line_by_receta = {line.receta_id: line for line in lineas}
        receta_ids = sorted(set(line_by_receta.keys()) | {p.receta_id for p in politicas})
        recetas_map = {
            r.id: r
            for r in Receta.objects.filter(id__in=receta_ids).select_related("rendimiento_unidad")
        }
        politicas_map = {p.receta_id: p for p in politicas}

        for rid in receta_ids:
            receta = recetas_map.get(rid)
            if not receta:
                continue
            linea = line_by_receta.get(rid)
            policy = politicas_map.get(rid)
            stock_reportado = _to_decimal_safe(linea.stock_reportado if linea else 0)
            en_transito = _to_decimal_safe(linea.en_transito if linea else 0)
            consumo_proyectado = _to_decimal_safe(linea.consumo_proyectado if linea else 0)
            sugerido_calc = _calcular_sugerido_reabasto(policy, stock_reportado, en_transito, consumo_proyectado)
            solicitado_val = _to_decimal_safe(linea.solicitado if linea else 0)
            rows_detalle.append(
                {
                    "receta": receta,
                    "linea": linea,
                    "policy": policy,
                    "stock_reportado": _quantize_qty(stock_reportado),
                    "en_transito": _quantize_qty(en_transito),
                    "consumo_proyectado": _quantize_qty(consumo_proyectado),
                    "sugerido": _quantize_qty(sugerido_calc),
                    "solicitado": _quantize_qty(solicitado_val),
                    "delta": _quantize_qty(solicitado_val - sugerido_calc),
                }
            )

        for policy in politicas:
            receta = recetas_map.get(policy.receta_id) or policy.receta
            if not receta:
                continue
            linea = line_by_receta.get(policy.receta_id)
            stock_reportado = _quantize_qty(_to_decimal_safe(linea.stock_reportado if linea else 0))
            sugerido = _quantize_qty(_calcular_sugerido_reabasto(policy, stock_reportado, Decimal("0"), Decimal("0")))
            solicitado_existente = _quantize_qty(_to_decimal_safe(linea.solicitado if linea else sugerido))
            rows_cierre.append(
                {
                    "receta": receta,
                    "policy": policy,
                    "linea": linea,
                    "stock_reportado": stock_reportado,
                    "sugerido": sugerido,
                    "solicitado": solicitado_existente,
                    "override_solicitado": bool(linea and solicitado_existente != sugerido),
                    "observaciones": (linea.observaciones if linea else ""),
                }
            )

    return {
        "solicitud": solicitud,
        "lineas": lineas,
        "politicas": politicas,
        "rows_detalle": rows_detalle,
        "rows_cierre": rows_cierre,
        "recetas_map": recetas_map,
    }


@login_required
def reabasto_cedis_captura(request: HttpRequest) -> HttpResponse:
    if not can_view_recetas(request.user):
        raise PermissionDenied("No tienes permisos para ver reabasto CEDIS.")

    if not is_branch_capture_only(request.user):
        fecha_operacion = _parse_date_safe(request.GET.get("fecha")) or timezone.localdate()
        sucursal_id = _to_int_safe(request.GET.get("sucursal_id"))
        return redirect(_reabasto_redirect(fecha_operacion, sucursal_id))

    fecha_operacion = _parse_date_safe(request.GET.get("fecha")) or (timezone.localdate() + timedelta(days=1))
    sucursal = _user_sucursal_scope(request.user)
    if not sucursal:
        messages.error(request, "Tu usuario no tiene sucursal asignada. Contacta a administración.")
        return render(
            request,
            "recetas/reabasto_cedis_captura.html",
            {
                "fecha_operacion": fecha_operacion,
                "sucursal": None,
                "solicitud": None,
                "rows_cierre": [],
                "erp_command_center": {
                    "owner": "Administración / DG",
                    "status": "Crítico",
                    "tone": "danger",
                    "blockers": 1,
                    "next_step": "Asignar sucursal al usuario para habilitar la captura diaria de cierre.",
                    "url": reverse("core:usuarios_accesos"),
                    "cta": "Abrir usuarios y accesos",
                },
                "erp_governance_rows": [],
                "executive_radar_rows": [],
                "critical_path_rows": [],
            },
        )

    data = _build_reabasto_rows(fecha_operacion, sucursal)
    rows_cierre = data["rows_cierre"]
    total_rows = len(rows_cierre)
    rows_pending = sum(1 for row in rows_cierre if Decimal(str(row.get("sugerido") or 0)) > 0)
    if total_rows == 0:
        erp_command_center = {
            "owner": f"Sucursal / {sucursal.codigo}",
            "status": "Crítico",
            "tone": "danger",
            "blockers": 1,
            "next_step": "Cargar políticas de stock mínimo para habilitar la captura de cierre y el cálculo del reabasto.",
            "url": reverse("recetas:reabasto_cedis"),
            "cta": "Abrir reabasto operativo",
        }
    elif rows_pending > 0:
        erp_command_center = {
            "owner": f"Sucursal / {sucursal.codigo}",
            "status": "En revisión",
            "tone": "warning",
            "blockers": rows_pending,
            "next_step": "Capturar stock final faltante y enviar el cierre para que CEDIS reciba la demanda completa.",
            "url": reverse("recetas:reabasto_cedis_captura"),
            "cta": "Completar captura",
        }
    else:
        erp_command_center = {
            "owner": f"Sucursal / {sucursal.codigo}",
            "status": "Estable",
            "tone": "success",
            "blockers": 0,
            "next_step": "El cierre diario está listo. Mantén el envío a CEDIS en tiempo para sostener producción y surtido.",
            "url": reverse("recetas:reabasto_cedis"),
            "cta": "Abrir reabasto CEDIS",
        }
    erp_governance_rows = [
        {
            "front": "Captura sucursal",
            "owner": f"Sucursal / {sucursal.codigo}",
            "blockers": rows_pending,
            "completion": 100 if rows_pending == 0 and total_rows > 0 else (70 if total_rows > 0 else 25),
            "detail": (
                "El cierre ya quedó capturado y listo para envío."
                if rows_pending == 0 and total_rows > 0
                else f"{rows_pending} producto(s) todavía requieren confirmación para cerrar el corte."
            ),
            "next_step": (
                "Mantener el cierre al día."
                if rows_pending == 0 and total_rows > 0
                else "Capturar stock final faltante y enviar a CEDIS."
            ),
            "url": reverse("recetas:reabasto_cedis_captura"),
            "cta": "Abrir captura",
        },
        {
            "front": "Reabasto CEDIS",
            "owner": "CEDIS / Producción",
            "blockers": rows_pending,
            "completion": 100 if rows_pending == 0 and total_rows > 0 else 80,
            "detail": (
                "CEDIS puede recibir el cierre sin brechas abiertas."
                if rows_pending == 0 and total_rows > 0
                else "El reabasto depende de completar el cierre de sucursal."
            ),
            "next_step": (
                "Continuar al plan de abastecimiento."
                if rows_pending == 0 and total_rows > 0
                else "Enviar la solicitud diaria a CEDIS."
            ),
            "url": reverse("recetas:reabasto_cedis"),
            "cta": "Abrir reabasto",
        },
    ]
    return render(
        request,
        "recetas/reabasto_cedis_captura.html",
        {
            "fecha_operacion": fecha_operacion,
            "sucursal": sucursal,
            "solicitud": data["solicitud"],
            "rows_cierre": rows_cierre,
            "erp_command_center": erp_command_center,
            "erp_governance_rows": erp_governance_rows,
            "executive_radar_rows": _recipes_executive_radar_rows(
                erp_governance_rows,
                owner=f"Sucursal / {sucursal.codigo}",
                fallback_url=reverse("recetas:reabasto_cedis_captura"),
            ),
            "critical_path_rows": _recipes_critical_path_rows(
                erp_governance_rows,
                owner=f"Sucursal / {sucursal.codigo}",
                fallback_url=reverse("recetas:reabasto_cedis_captura"),
            ),
        },
    )


@login_required
def reabasto_cedis(request: HttpRequest) -> HttpResponse:
    if not can_view_recetas(request.user):
        raise PermissionDenied("No tienes permisos para ver reabasto CEDIS.")

    branch_capture_only_mode = is_branch_capture_only(request.user)
    if branch_capture_only_mode:
        fecha_operacion = _parse_date_safe(request.GET.get("fecha")) or (timezone.localdate() + timedelta(days=1))
        user_sucursal = _user_sucursal_scope(request.user)
        sucursal_id = user_sucursal.id if user_sucursal else None
        return redirect(_reabasto_redirect(fecha_operacion, sucursal_id, capture_only=True))
    user_sucursal = _user_sucursal_scope(request.user)
    has_sucursal_scope = bool(user_sucursal)
    # Si un usuario tiene sucursal asignada, este módulo opera en modo captura simple para evitar flujo avanzado.
    sucursal_locked = bool(branch_capture_only_mode or has_sucursal_scope)
    can_manage = can_manage_compras(request.user) and not sucursal_locked

    fecha_raw = request.GET.get("fecha")
    if _parse_date_safe(fecha_raw):
        fecha_operacion = _parse_date_safe(fecha_raw)
    else:
        fecha_operacion = (timezone.localdate() + timedelta(days=1)) if sucursal_locked else timezone.localdate()

    sucursales = list(sucursales_operativas().order_by("codigo"))
    if sucursal_locked and user_sucursal and all(s.id != user_sucursal.id for s in sucursales):
        sucursales = [user_sucursal] + sucursales

    sucursal_id = _to_int_safe(request.GET.get("sucursal_id"))
    sucursal = None
    if sucursal_locked and user_sucursal:
        sucursal = user_sucursal
    elif sucursal_locked and not user_sucursal:
        messages.error(request, "Este usuario de sucursal no tiene sucursal asignada. Contacta a administración.")
        sucursal = None
    elif sucursal_id:
        sucursal = next((s for s in sucursales if s.id == sucursal_id), None)
    if not sucursal and sucursales and not sucursal_locked and sucursal_id:
        sucursal = sucursales[0]

    data = _build_reabasto_rows(fecha_operacion, sucursal)
    solicitud = data["solicitud"]
    politicas = data["politicas"]
    rows_detalle = data["rows_detalle"]
    rows_cierre = data["rows_cierre"]

    recetas_producto = list(
        Receta.objects.filter(tipo=Receta.TIPO_PRODUCTO_FINAL)
        .order_by("nombre")
        .only("id", "nombre", "codigo_point")
    )
    if not recetas_producto:
        recetas_producto = list(Receta.objects.order_by("nombre").only("id", "nombre", "codigo_point")[:400])

    inventario_cedis = list(
        InventarioCedisProducto.objects.select_related("receta")
        .filter(receta__in=[r.id for r in recetas_producto])
        .order_by("receta__nombre")
    )
    inventario_map = {inv.receta_id: inv for inv in inventario_cedis}
    productos_sin_inventario = [r for r in recetas_producto if r.id not in inventario_map]
    reabasto_enterprise_board = _reabasto_enterprise_board(
        recetas_producto=recetas_producto,
        inventario_map=inventario_map,
        fecha_operacion=fecha_operacion,
    )

    consolidado_rows = _consolidado_reabasto_por_fecha(fecha_operacion)
    total_solicitado = sum((row["total_solicitado"] for row in consolidado_rows), Decimal("0"))
    total_faltante = sum((row["cedis_faltante_producir"] for row in consolidado_rows), Decimal("0"))
    stage_key = (request.GET.get("stage_key") or "auto").strip().lower()
    closure_key = (request.GET.get("closure_key") or "auto").strip().lower()
    handoff_key = (request.GET.get("handoff_key") or "auto").strip().lower()
    master_focus_key = (request.GET.get("master_focus_key") or "auto").strip()
    board_focus_key = (request.GET.get("board_focus_key") or "auto").strip().lower()

    solicitudes_hoy = list(
        SolicitudReabastoCedis.objects.filter(fecha_operacion=fecha_operacion)
        .select_related("sucursal")
        .order_by("sucursal__codigo")
    )
    resumen_cierre = _resumen_cierre_sucursales_reabasto(fecha_operacion, sucursales)
    reabasto_daily_control = _reabasto_daily_control(
        fecha_operacion=fecha_operacion,
        resumen_cierre=resumen_cierre,
        reabasto_enterprise_board=reabasto_enterprise_board,
        consolidado_rows=consolidado_rows,
        stage_key=stage_key,
        closure_key=closure_key,
        handoff_key=handoff_key,
        master_focus_key=master_focus_key,
    )
    demand_history_summary = _reabasto_demand_history_summary(fecha_operacion)
    daily_decision_rows = _reabasto_daily_decisions(
        demand_history_summary=demand_history_summary,
        daily_control=reabasto_daily_control,
    )
    branch_priority_rows = _reabasto_branch_priority_rows(
        fecha_operacion=fecha_operacion,
        sucursales=sucursales,
        resumen_cierre=resumen_cierre,
        demand_history_summary=demand_history_summary,
    )
    branch_supply_rows = _reabasto_branch_supply_rows(
        fecha_operacion=fecha_operacion,
        branch_priority_rows=branch_priority_rows,
    )

    board_cards = list(reabasto_enterprise_board.get("blocker_cards") or [])
    valid_board_focus_keys = {str(card.get("key") or "").strip().lower() for card in board_cards if card.get("key")}
    selected_board_focus_key = board_focus_key if board_focus_key in valid_board_focus_keys else "auto"
    board_focus_base = reverse("recetas:reabasto_cedis") + f"?{urlencode({'fecha': fecha_operacion.isoformat()})}"
    for card in board_cards:
        card_key = str(card.get("key") or "").strip().lower()
        card["focus_url"] = f"{board_focus_base}&board_focus_key={urlencode({'k': card_key})[2:]}"
        card["is_active"] = selected_board_focus_key != "auto" and card_key == selected_board_focus_key
    board_detail_source = list(reabasto_enterprise_board.get("detail_rows") or [])
    board_detail_rows = (
        [row for row in board_detail_source if str(row.get("blocker_key") or "").strip().lower() == selected_board_focus_key]
        if selected_board_focus_key != "auto"
        else board_detail_source
    )
    board_focus = None
    if board_detail_rows:
        first_board_focus = board_detail_rows[0]
        board_focus = {
            "label": first_board_focus.get("blocker_label", "Bloqueo operativo"),
            "summary": (
                f"El abastecimiento CEDIS sigue condicionado por {first_board_focus.get('receta_nombre', 'una receta')} "
                f"en el bloqueo {first_board_focus.get('blocker_label', '').lower()}."
            ),
            "action_label": first_board_focus.get("action_label", "Abrir detalle"),
            "action_url": first_board_focus.get("action_url", "#"),
        "tone": "danger" if first_board_focus.get("blocker_label") in {"Sin inventario CEDIS", "Receta por validar"} else "warning",
        }
    reabasto_enterprise_board["blocker_cards"] = board_cards
    reabasto_enterprise_board["detail_rows"] = board_detail_rows[:12]
    reabasto_enterprise_board["selected_focus_key"] = selected_board_focus_key
    reabasto_enterprise_board["focus"] = board_focus

    critical_path_source = list(reabasto_daily_control.get("document_stage_rows") or [])
    if critical_path_source:
        critical_path_rows = _recipes_critical_path_rows(
            critical_path_source,
            owner="Sucursal / CEDIS",
            fallback_url=reverse("recetas:reabasto_cedis"),
        )
        executive_radar_rows = _recipes_executive_radar_rows(
            critical_path_source,
            owner="Sucursal / CEDIS",
            fallback_url=reverse("recetas:reabasto_cedis"),
        )
    else:
        executive_source = list(reabasto_daily_control.get("control_cards") or [])
        critical_path_rows = _critical_path_rows_from_cards(
            executive_source,
            owner="Sucursal / CEDIS",
            fallback_url=reverse("recetas:reabasto_cedis"),
            default_dependency="Cierre operativo de sucursales",
        )
        executive_radar_rows = _recipes_executive_radar_rows(
            executive_source,
            owner="Sucursal / CEDIS",
            fallback_url=reverse("recetas:reabasto_cedis"),
        )

    return render(
        request,
        "recetas/reabasto_cedis.html",
        {
            "can_manage_reabasto": can_manage,
            "sucursal_locked": sucursal_locked,
            "user_sucursal": user_sucursal,
            "branch_capture_only_mode": branch_capture_only_mode,
            "show_simple_capture": True,
            "fecha_operacion": fecha_operacion,
            "sucursales": sucursales,
            "sucursal": sucursal,
            "solicitud": solicitud,
            "solicitudes_hoy": solicitudes_hoy,
            "rows_detalle": rows_detalle,
            "rows_cierre": rows_cierre,
            "recetas_producto": recetas_producto,
            "inventario_cedis": inventario_cedis,
            "productos_sin_inventario": productos_sin_inventario,
            "reabasto_enterprise_board": reabasto_enterprise_board,
            "consolidado_rows": consolidado_rows,
            "total_solicitado": _quantize_qty(total_solicitado),
            "total_faltante": _quantize_qty(total_faltante),
            "resumen_cierre": resumen_cierre,
            "reabasto_daily_control": reabasto_daily_control,
            "demand_history_summary": demand_history_summary,
            "daily_decision_rows": daily_decision_rows,
            "branch_priority_rows": branch_priority_rows,
            "branch_supply_rows": branch_supply_rows,
            "executive_radar_rows": executive_radar_rows,
            "critical_path_rows": critical_path_rows,
        },
    )


@login_required
@require_POST
def reabasto_cedis_politica_guardar(request: HttpRequest) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para editar políticas de stock.")

    fecha_operacion = _parse_date_safe(request.POST.get("fecha_operacion")) or timezone.localdate()
    sucursal = get_object_or_404(Sucursal, pk=_to_int_safe(request.POST.get("sucursal_id")))
    receta = get_object_or_404(Receta, pk=_to_int_safe(request.POST.get("receta_id")))

    policy, _ = PoliticaStockSucursalProducto.objects.get_or_create(
        sucursal=sucursal,
        receta=receta,
        defaults={"activa": True},
    )
    policy.stock_minimo = _quantize_qty(_to_decimal_safe(request.POST.get("stock_minimo")))
    policy.stock_objetivo = _quantize_qty(_to_decimal_safe(request.POST.get("stock_objetivo")))
    policy.stock_maximo = _quantize_qty(_to_decimal_safe(request.POST.get("stock_maximo")))
    policy.dias_cobertura = max(1, _to_int_safe(request.POST.get("dias_cobertura"), 1))
    policy.stock_seguridad = _quantize_qty(_to_decimal_safe(request.POST.get("stock_seguridad")))
    policy.lote_minimo = _quantize_qty(_to_decimal_safe(request.POST.get("lote_minimo")))
    policy.multiplo_empaque = _quantize_qty(_to_decimal_safe(request.POST.get("multiplo_empaque"))) or Decimal("1")
    policy.activa = request.POST.get("activa", "on") == "on"
    policy.save()

    messages.success(request, "Política de stock guardada.")
    return redirect(_reabasto_redirect(fecha_operacion, sucursal.id))


@login_required
@require_POST
def reabasto_cedis_inventario_guardar(request: HttpRequest) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para editar inventario CEDIS.")

    fecha_operacion = _parse_date_safe(request.POST.get("fecha_operacion")) or timezone.localdate()
    sucursal_id = _to_int_safe(request.POST.get("sucursal_id"))
    receta = get_object_or_404(Receta, pk=_to_int_safe(request.POST.get("receta_id")))

    inv, _ = InventarioCedisProducto.objects.get_or_create(receta=receta)
    inv.stock_actual = _quantize_qty(_to_decimal_safe(request.POST.get("stock_actual")))
    inv.stock_reservado = _quantize_qty(_to_decimal_safe(request.POST.get("stock_reservado")))
    inv.save()

    messages.success(request, "Inventario CEDIS actualizado.")
    return redirect(_reabasto_redirect(fecha_operacion, sucursal_id or None))


@login_required
@require_POST
def reabasto_cedis_linea_guardar(request: HttpRequest) -> HttpResponse:
    if not can_view_recetas(request.user):
        raise PermissionDenied("No tienes permisos para registrar solicitud de reabasto.")

    fecha_operacion = _parse_date_safe(request.POST.get("fecha_operacion")) or timezone.localdate()
    sucursal = get_object_or_404(Sucursal, pk=_to_int_safe(request.POST.get("sucursal_id")))
    _assert_can_capture_for_sucursal(request.user, sucursal)
    receta = get_object_or_404(Receta, pk=_to_int_safe(request.POST.get("receta_id")))
    solicitud, _ = SolicitudReabastoCedis.objects.get_or_create(
        fecha_operacion=fecha_operacion,
        sucursal=sucursal,
        defaults={"creado_por": request.user},
    )

    stock_reportado = _quantize_qty(_to_decimal_safe(request.POST.get("stock_reportado")))
    en_transito = _quantize_qty(_to_decimal_safe(request.POST.get("en_transito")))
    consumo_proyectado = _quantize_qty(_to_decimal_safe(request.POST.get("consumo_proyectado")))
    solicitado_raw = (request.POST.get("solicitado") or "").strip()

    policy = (
        PoliticaStockSucursalProducto.objects.filter(
            sucursal=sucursal,
            receta=receta,
            activa=True,
        )
        .order_by("-id")
        .first()
    )
    sugerido = _calcular_sugerido_reabasto(policy, stock_reportado, en_transito, consumo_proyectado)
    if solicitado_raw == "":
        solicitado = sugerido
    else:
        solicitado = _quantize_qty(_to_decimal_safe(solicitado_raw))
        solicitado = _aplicar_lote_multiplo(solicitado, policy)

    linea, created = SolicitudReabastoCedisLinea.objects.get_or_create(
        solicitud=solicitud,
        receta=receta,
        defaults={
            "stock_reportado": stock_reportado,
            "en_transito": en_transito,
            "consumo_proyectado": consumo_proyectado,
            "sugerido": sugerido,
            "solicitado": solicitado,
            "justificacion": (request.POST.get("justificacion") or "").strip()[:255],
            "observaciones": (request.POST.get("observaciones") or "").strip()[:255],
        },
    )
    if not created:
        linea.stock_reportado = stock_reportado
        linea.en_transito = en_transito
        linea.consumo_proyectado = consumo_proyectado
        linea.sugerido = sugerido
        linea.solicitado = solicitado
        linea.justificacion = (request.POST.get("justificacion") or "").strip()[:255]
        linea.observaciones = (request.POST.get("observaciones") or "").strip()[:255]
        linea.save()

    if solicitud.estado == SolicitudReabastoCedis.ESTADO_CANCELADA:
        solicitud.estado = SolicitudReabastoCedis.ESTADO_BORRADOR
        solicitud.save(update_fields=["estado", "actualizado_en"])

    messages.success(request, "Línea de reabasto guardada.")
    return redirect(_reabasto_redirect(fecha_operacion, sucursal.id))


@login_required
@require_POST
def reabasto_cedis_cierre_guardar(request: HttpRequest) -> HttpResponse:
    if not can_view_recetas(request.user):
        raise PermissionDenied("No tienes permisos para registrar cierre de sucursal.")

    fecha_operacion = _parse_date_safe(request.POST.get("fecha_operacion")) or timezone.localdate()
    sucursal = get_object_or_404(Sucursal, pk=_to_int_safe(request.POST.get("sucursal_id")))
    capture_only_mode = is_branch_capture_only(request.user)
    _assert_can_capture_for_sucursal(request.user, sucursal)

    solicitud, _ = SolicitudReabastoCedis.objects.get_or_create(
        fecha_operacion=fecha_operacion,
        sucursal=sucursal,
        defaults={"creado_por": request.user},
    )

    receta_ids = []
    for rid in request.POST.getlist("row_receta_id"):
        rid_int = _to_int_safe(rid)
        if rid_int and rid_int not in receta_ids:
            receta_ids.append(rid_int)
    if not receta_ids:
        messages.error(request, "No se recibieron productos de cierre para procesar.")
        return redirect(_reabasto_redirect(fecha_operacion, sucursal.id, capture_only=capture_only_mode))

    receta_map = Receta.objects.in_bulk(receta_ids)
    policy_map = {
        p.receta_id: p
        for p in PoliticaStockSucursalProducto.objects.filter(
            sucursal=sucursal,
            activa=True,
            receta_id__in=receta_ids,
        ).only("id", "receta_id", "stock_minimo", "stock_objetivo", "stock_seguridad", "lote_minimo", "multiplo_empaque")
    }
    existing_map = {
        line.receta_id: line
        for line in SolicitudReabastoCedisLinea.objects.filter(
            solicitud=solicitud,
            receta_id__in=receta_ids,
        )
    }

    created = 0
    updated = 0
    skipped = 0
    for receta_id in receta_ids:
        receta = receta_map.get(receta_id)
        if not receta:
            skipped += 1
            continue

        line = existing_map.get(receta_id)
        stock_raw = (request.POST.get(f"stock_reportado_{receta_id}") or "").strip()
        solicitado_raw = (request.POST.get(f"solicitado_{receta_id}") or "").strip()
        observaciones = (request.POST.get(f"observaciones_{receta_id}") or "").strip()[:255]

        if not stock_raw and not solicitado_raw and not line:
            skipped += 1
            continue

        if stock_raw:
            stock_reportado = _quantize_qty(_to_decimal_safe(stock_raw))
        else:
            stock_reportado = _quantize_qty(_to_decimal_safe(line.stock_reportado if line else 0))

        policy = policy_map.get(receta_id)
        sugerido = _calcular_sugerido_reabasto(policy, stock_reportado, Decimal("0"), Decimal("0"))

        if solicitado_raw:
            solicitado = _quantize_qty(_to_decimal_safe(solicitado_raw))
            solicitado = _aplicar_lote_multiplo(solicitado, policy)
        else:
            solicitado = _quantize_qty(sugerido)

        defaults = {
            "stock_reportado": stock_reportado,
            "en_transito": Decimal("0"),
            "consumo_proyectado": Decimal("0"),
            "sugerido": _quantize_qty(sugerido),
            "solicitado": solicitado,
            "justificacion": "Cierre sucursal",
            "observaciones": observaciones,
        }
        if line:
            line.stock_reportado = defaults["stock_reportado"]
            line.en_transito = defaults["en_transito"]
            line.consumo_proyectado = defaults["consumo_proyectado"]
            line.sugerido = defaults["sugerido"]
            line.solicitado = defaults["solicitado"]
            line.justificacion = defaults["justificacion"]
            line.observaciones = defaults["observaciones"]
            line.save()
            updated += 1
        else:
            SolicitudReabastoCedisLinea.objects.create(
                solicitud=solicitud,
                receta=receta,
                **defaults,
            )
            created += 1

    accion = (request.POST.get("accion") or "BORRADOR").strip().upper()
    if accion == "ENVIAR":
        solicitud.estado = SolicitudReabastoCedis.ESTADO_ENVIADA
        msg_estado = "enviada a CEDIS"
    else:
        if solicitud.estado != SolicitudReabastoCedis.ESTADO_ATENDIDA:
            solicitud.estado = SolicitudReabastoCedis.ESTADO_BORRADOR
        msg_estado = "guardada en borrador"

    notas = (request.POST.get("notas_cierre") or "").strip()
    if notas:
        solicitud.notas = notas[:255]
    solicitud.save(update_fields=["estado", "notas", "actualizado_en"])

    messages.success(
        request,
        f"Cierre {msg_estado}. Componentes creados: {created}, actualizados: {updated}, omitidos: {skipped}.",
    )
    return redirect(_reabasto_redirect(fecha_operacion, sucursal.id, capture_only=capture_only_mode))


@login_required
@require_POST
def reabasto_cedis_linea_eliminar(request: HttpRequest, linea_id: int) -> HttpResponse:
    if not can_view_recetas(request.user):
        raise PermissionDenied("No tienes permisos para modificar solicitud de reabasto.")

    linea = get_object_or_404(
        SolicitudReabastoCedisLinea.objects.select_related("solicitud", "solicitud__sucursal"),
        pk=linea_id,
    )
    _assert_can_capture_for_sucursal(request.user, linea.solicitud.sucursal)
    fecha_operacion = linea.solicitud.fecha_operacion
    sucursal_id = linea.solicitud.sucursal_id
    linea.delete()
    messages.success(request, "Línea eliminada.")
    return redirect(_reabasto_redirect(fecha_operacion, sucursal_id))


@login_required
@require_POST
def reabasto_cedis_estado_guardar(request: HttpRequest, solicitud_id: int) -> HttpResponse:
    solicitud = get_object_or_404(SolicitudReabastoCedis, pk=solicitud_id)
    _assert_can_capture_for_sucursal(request.user, solicitud.sucursal)
    estado = (request.POST.get("estado") or "").strip().upper()
    if estado not in {
        SolicitudReabastoCedis.ESTADO_BORRADOR,
        SolicitudReabastoCedis.ESTADO_ENVIADA,
        SolicitudReabastoCedis.ESTADO_ATENDIDA,
        SolicitudReabastoCedis.ESTADO_CANCELADA,
    }:
        messages.error(request, "Estado inválido.")
        return redirect(_reabasto_redirect(solicitud.fecha_operacion, solicitud.sucursal_id))

    if estado in {SolicitudReabastoCedis.ESTADO_ATENDIDA, SolicitudReabastoCedis.ESTADO_CANCELADA} and not can_manage_compras(
        request.user
    ):
        raise PermissionDenied("Solo compras/administración puede cerrar o cancelar solicitudes.")

    solicitud.estado = estado
    solicitud.notas = (request.POST.get("notas") or solicitud.notas or "").strip()
    solicitud.save(update_fields=["estado", "notas", "actualizado_en"])
    messages.success(request, "Estado de solicitud actualizado.")
    return redirect(_reabasto_redirect(solicitud.fecha_operacion, solicitud.sucursal_id))


@login_required
@require_POST
def reabasto_cedis_importar(request: HttpRequest) -> HttpResponse:
    if not can_view_recetas(request.user):
        raise PermissionDenied("No tienes permisos para importar solicitudes de reabasto.")

    fecha_operacion = _parse_date_safe(request.POST.get("fecha_operacion")) or timezone.localdate()
    sucursal = get_object_or_404(Sucursal, pk=_to_int_safe(request.POST.get("sucursal_id")))
    _assert_can_capture_for_sucursal(request.user, sucursal)
    uploaded = request.FILES.get("archivo")
    if not uploaded:
        messages.error(request, "Selecciona un archivo CSV/XLSX para importar.")
        return redirect(_reabasto_redirect(fecha_operacion, sucursal.id))

    try:
        rows = _load_reabasto_rows(uploaded)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect(_reabasto_redirect(fecha_operacion, sucursal.id))

    if not rows:
        messages.warning(request, "El archivo no contiene renglones.")
        return redirect(_reabasto_redirect(fecha_operacion, sucursal.id))

    solicitud, _ = SolicitudReabastoCedis.objects.get_or_create(
        fecha_operacion=fecha_operacion,
        sucursal=sucursal,
        defaults={"creado_por": request.user},
    )

    created = 0
    updated = 0
    skipped = 0
    for row in rows:
        receta = _resolve_receta_reabasto(str(row.get("receta") or ""), str(row.get("codigo_point") or ""))
        if not receta:
            skipped += 1
            continue

        policy = (
            PoliticaStockSucursalProducto.objects.filter(
                sucursal=sucursal,
                receta=receta,
                activa=True,
            )
            .order_by("-id")
            .first()
        )
        stock_reportado = _quantize_qty(_to_decimal_safe(row.get("stock_reportado")))
        en_transito = _quantize_qty(_to_decimal_safe(row.get("en_transito")))
        consumo_proyectado = _quantize_qty(_to_decimal_safe(row.get("consumo_proyectado")))
        sugerido = _calcular_sugerido_reabasto(policy, stock_reportado, en_transito, consumo_proyectado)

        solicitado_raw = str(row.get("solicitado") or "").strip()
        if solicitado_raw == "":
            solicitado = sugerido
        else:
            solicitado = _quantize_qty(_to_decimal_safe(solicitado_raw))
            solicitado = _aplicar_lote_multiplo(solicitado, policy)

        linea, was_created = SolicitudReabastoCedisLinea.objects.get_or_create(
            solicitud=solicitud,
            receta=receta,
            defaults={
                "stock_reportado": stock_reportado,
                "en_transito": en_transito,
                "consumo_proyectado": consumo_proyectado,
                "sugerido": sugerido,
                "solicitado": solicitado,
                "justificacion": (str(row.get("justificacion") or "").strip())[:255],
                "observaciones": (str(row.get("observaciones") or "").strip())[:255],
            },
        )
        if was_created:
            created += 1
        else:
            linea.stock_reportado = stock_reportado
            linea.en_transito = en_transito
            linea.consumo_proyectado = consumo_proyectado
            linea.sugerido = sugerido
            linea.solicitado = solicitado
            linea.justificacion = (str(row.get("justificacion") or "").strip())[:255]
            linea.observaciones = (str(row.get("observaciones") or "").strip())[:255]
            linea.save()
            updated += 1

    messages.success(
        request,
        f"Importación lista. Creadas: {created}, actualizadas: {updated}, omitidas: {skipped}.",
    )
    return redirect(_reabasto_redirect(fecha_operacion, sucursal.id))


@login_required
def reabasto_cedis_consolidado_export(request: HttpRequest) -> HttpResponse:
    if not can_view_recetas(request.user):
        raise PermissionDenied("No tienes permisos para exportar consolidado CEDIS.")

    fecha_operacion = _parse_date_safe(request.GET.get("fecha")) or timezone.localdate()
    export_format = (request.GET.get("format") or "xlsx").strip().lower()
    rows = _consolidado_reabasto_por_fecha(fecha_operacion)

    if export_format == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="cedis_consolidado_{fecha_operacion.isoformat()}.csv"'
        writer = csv.writer(response)
        writer.writerow(
            [
                "receta",
                "sucursales",
                "total_sugerido",
                "total_solicitado",
                "cedis_disponible",
                "cedis_faltante_producir",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row["receta"],
                    row["sucursales"],
                    float(row["total_sugerido"]),
                    float(row["total_solicitado"]),
                    float(row["cedis_disponible"]),
                    float(row["cedis_faltante_producir"]),
                ]
            )
        return response

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado CEDIS"
    ws.append(
        [
            "Receta",
            "Sucursales",
            "Total sugerido",
            "Total solicitado",
            "CEDIS disponible",
            "CEDIS faltante producir",
        ]
    )
    for row in rows:
        ws.append(
            [
                row["receta"],
                row["sucursales"],
                float(row["total_sugerido"]),
                float(row["total_solicitado"]),
                float(row["cedis_disponible"]),
                float(row["cedis_faltante_producir"]),
            ]
        )
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="cedis_consolidado_{fecha_operacion.isoformat()}.xlsx"'
    return response


@login_required
@require_POST
def reabasto_cedis_generar_plan(request: HttpRequest) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para generar plan desde reabasto CEDIS.")

    fecha_operacion = _parse_date_safe(request.POST.get("fecha_operacion")) or timezone.localdate()
    resumen_cierre = _resumen_cierre_sucursales_reabasto(
        fecha_operacion,
        list(sucursales_operativas().order_by("codigo")),
    )
    enterprise_context = _reabasto_enterprise_context(fecha_operacion)
    reabasto_enterprise_board = enterprise_context["reabasto_enterprise_board"]
    consolidado_rows = _consolidado_reabasto_por_fecha(fecha_operacion)
    plan_existente = _find_reabasto_plan(fecha_operacion)
    doc_control = _plan_document_control(plan_existente) if plan_existente else None
    demand_history_summary = _reabasto_demand_history_summary(fecha_operacion)
    generation_gate = _reabasto_generation_gate(
        fecha_operacion=fecha_operacion,
        resumen_cierre=resumen_cierre,
        reabasto_enterprise_board=reabasto_enterprise_board,
        consolidado_rows=consolidado_rows,
        plan=plan_existente,
        doc_control=doc_control,
        demand_history_summary=demand_history_summary,
    )
    if not resumen_cierre["listo_8am"] and resumen_cierre["pendientes_codigos"]:
        messages.warning(
            request,
            f"Aviso operativo 8:00 AM: faltan o están tardías sucursales ({', '.join(resumen_cierre['pendientes_codigos'])}).",
        )
    if not generation_gate["can_generate_plan"]:
        _log_reabasto_gate_block(
            request.user,
            fecha_operacion=fecha_operacion,
            target="plan",
            generation_gate=generation_gate,
        )
        messages.warning(
            request,
            _reabasto_generation_blocker_message(generation_gate, target="plan"),
        )
        return redirect(_reabasto_redirect(fecha_operacion))
    plan, created_items, total_qty = _upsert_plan_reabasto_cedis(fecha_operacion, request.user)
    if not plan:
        messages.warning(
            request,
            "No hay faltantes a producir para esta fecha. Primero registra solicitudes de sucursal.",
        )
        return redirect(_reabasto_redirect(fecha_operacion))
    messages.success(
        request,
        f"Plan CEDIS generado/actualizado: {plan.nombre}. Renglones: {created_items}.",
    )
    return redirect(f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}")


def _upsert_plan_reabasto_cedis(fecha_operacion: date, user) -> tuple[PlanProduccion | None, int, Decimal]:
    consolidado_rows = _consolidado_reabasto_por_fecha(fecha_operacion)
    faltantes = [row for row in consolidado_rows if _to_decimal_safe(row.get("cedis_faltante_producir")) > 0]
    if not faltantes:
        return None, 0, Decimal("0")

    marker = f"[AUTO_REABASTO_CEDIS:{fecha_operacion.isoformat()}]"
    nombre_plan = f"CEDIS Reabasto {fecha_operacion.isoformat()}"
    plan = (
        PlanProduccion.objects.filter(
            fecha_produccion=fecha_operacion,
            nombre=nombre_plan,
        )
        .order_by("-id")
        .first()
    )
    if not plan:
        plan = PlanProduccion.objects.create(
            nombre=nombre_plan,
            fecha_produccion=fecha_operacion,
            notas=f"{marker} Generado automáticamente desde consolidado de reabasto CEDIS.",
            creado_por=user,
        )
    else:
        notas_actuales = (plan.notas or "").strip()
        if marker not in notas_actuales:
            plan.notas = f"{marker}\n{notas_actuales}".strip()
            plan.save(update_fields=["notas", "actualizado_en"])

    plan.items.all().delete()
    created_items = 0
    for row in faltantes:
        rid = _to_int_safe(row.get("receta_id"))
        if not rid:
            continue
        receta = Receta.objects.filter(pk=rid).first()
        if not receta:
            continue
        cantidad = _quantize_qty(_to_decimal_safe(row.get("cedis_faltante_producir")))
        if cantidad <= 0:
            continue
        PlanProduccionItem.objects.create(
            plan=plan,
            receta=receta,
            cantidad=cantidad,
            notas="AUTO_REABASTO_CEDIS",
        )
        created_items += 1

    if created_items <= 0:
        return None, 0, Decimal("0")

    total_qty = sum((_to_decimal_safe(row.get("cedis_faltante_producir")) for row in faltantes), Decimal("0"))
    log_event(
        user,
        "CREATE",
        "recetas.PlanProduccion",
        plan.id,
        {
            "source": "REABASTO_CEDIS",
            "fecha_operacion": fecha_operacion.isoformat(),
            "items": created_items,
            "qty_total": float(total_qty),
            "plan_nombre": plan.nombre,
        },
    )
    return plan, created_items, total_qty


@login_required
@require_POST
def reabasto_cedis_generar_compras(request: HttpRequest) -> HttpResponse:
    if not can_manage_compras(request.user):
        raise PermissionDenied("No tienes permisos para generar compras desde reabasto CEDIS.")

    fecha_operacion = _parse_date_safe(request.POST.get("fecha_operacion")) or timezone.localdate()
    resumen_cierre = _resumen_cierre_sucursales_reabasto(
        fecha_operacion,
        list(sucursales_operativas().order_by("codigo")),
    )
    enterprise_context = _reabasto_enterprise_context(fecha_operacion)
    reabasto_enterprise_board = enterprise_context["reabasto_enterprise_board"]
    consolidado_rows = _consolidado_reabasto_por_fecha(fecha_operacion)
    plan_existente = _find_reabasto_plan(fecha_operacion)
    doc_control = _plan_document_control(plan_existente) if plan_existente else None
    demand_history_summary = _reabasto_demand_history_summary(fecha_operacion)
    generation_gate = _reabasto_generation_gate(
        fecha_operacion=fecha_operacion,
        resumen_cierre=resumen_cierre,
        reabasto_enterprise_board=reabasto_enterprise_board,
        consolidado_rows=consolidado_rows,
        plan=plan_existente,
        doc_control=doc_control,
        demand_history_summary=demand_history_summary,
    )
    if not resumen_cierre["listo_8am"] and resumen_cierre["pendientes_codigos"]:
        messages.warning(
            request,
            f"Aviso operativo 8:00 AM: faltan o están tardías sucursales ({', '.join(resumen_cierre['pendientes_codigos'])}).",
        )
    if not generation_gate["can_generate_compras"]:
        _log_reabasto_gate_block(
            request.user,
            fecha_operacion=fecha_operacion,
            target="compras",
            generation_gate=generation_gate,
        )
        messages.warning(
            request,
            _reabasto_generation_blocker_message(generation_gate, target="compras"),
        )
        return redirect(_reabasto_redirect(fecha_operacion))
    plan, created_items, _ = _upsert_plan_reabasto_cedis(fecha_operacion, request.user)
    if not plan:
        messages.warning(
            request,
            "No hay faltantes a producir para generar compras en esta fecha.",
        )
        return redirect(_reabasto_redirect(fecha_operacion))

    stats = _generar_solicitudes_compra_desde_plan(
        plan=plan,
        user=request.user,
        replace_prev=True,
        auto_create_oc=True,
    )

    if stats["creadas"] == 0 and stats["actualizadas"] == 0:
        messages.warning(
            request,
            f"Plan {plan.nombre} (renglones {created_items}) generado, pero no hubo materia prima para solicitudes de compra.",
        )
    else:
        messages.success(
            request,
            (
                f"Plan {plan.nombre} listo ({created_items} renglones). "
                f"Solicitudes compra: creadas {stats['creadas']}, actualizadas {stats['actualizadas']}. "
                f"OC borrador por proveedor: creadas {stats['oc_creadas']}, actualizadas {stats['oc_actualizadas']}."
            ),
        )
    compras_query = urlencode(
        {
            "source": "plan",
            "plan_id": str(plan.id),
            "reabasto": "all",
        }
    )
    return redirect(f"{reverse('compras:solicitudes')}?{compras_query}")


@login_required
def mrp_form(request: HttpRequest) -> HttpResponse:
    recetas = Receta.objects.order_by("nombre")
    resultado = None
    forecast_preview = request.session.get("pronostico_estadistico_preview")
    forecast_compare_escenario = str((forecast_preview or {}).get("escenario") or "base").strip().lower()
    if forecast_compare_escenario not in {"base", "bajo", "alto"}:
        forecast_compare_escenario = "base"
    try:
        forecast_vs_solicitud = _forecast_vs_solicitud_preview(
            forecast_preview,
            escenario=forecast_compare_escenario,
        )
    except (OperationalError, ProgrammingError):
        forecast_vs_solicitud = None
    erp_command_center = {
        "owner": "Producción / Costeo",
        "status": "Listo",
        "tone": "success",
        "blockers": 0,
        "next_step": "Selecciona una receta y ejecuta el cálculo para revisar estructura BOM, maestro y capacidad antes de liberar compras o producción.",
        "url": reverse("recetas:recetas_list"),
        "cta": "Abrir catálogo de recetas",
    }
    workflow_rows = [
        {
            "step": "01",
            "title": "Estructura BOM",
            "owner": "Producción / Costeo",
            "open": 0,
            "closed": 1,
            "completion": 100,
            "tone": "success",
            "detail": "La receta debe tener cantidades, derivados y empaques consistentes antes del cálculo.",
            "next_step": "Validar estructura base del documento.",
            "action_label": "Abrir recetas",
            "action_href": reverse("recetas:recetas_list"),
        },
        {
            "step": "02",
            "title": "Maestro del artículo",
            "owner": "Maestros / DG",
            "open": 0,
            "closed": 1,
            "completion": 100,
            "tone": "success",
            "detail": "Cada componente debe usar artículo maestro, unidad y costo vigente.",
            "next_step": "Corregir artículos incompletos antes de comprar o producir.",
            "action_label": "Abrir maestro",
            "action_href": reverse("maestros:insumo_list"),
        },
        {
            "step": "03",
            "title": "Capacidad operativa",
            "owner": "Inventario / Compras",
            "open": 0,
            "closed": 1,
            "completion": 100,
            "tone": "success",
            "detail": "El cálculo debe confirmar cobertura de stock y siguiente acción operativa.",
            "next_step": "Usar el resultado para plan, compra o producción.",
            "action_label": "Abrir existencias",
            "action_href": reverse("inventario:existencias"),
        },
    ]
    if request.method == "POST":
        receta_id = request.POST.get("receta_id")
        mult = request.POST.get("multiplicador", "1").strip()
        focus_kind = (request.POST.get("focus_kind") or "").strip().lower()
        focus_key = (request.POST.get("focus_key") or "").strip()
        try:
            multiplicador = Decimal(mult)
        except Exception:
            multiplicador = Decimal("1")

        receta = get_object_or_404(Receta, pk=receta_id)
        try:
            demand_signal = _mrp_recipe_demand_signal(
                receta,
                forecast_preview,
                forecast_vs_solicitud,
            )
        except (OperationalError, ProgrammingError):
            demand_signal = None
        commercial_gate = _commercial_signal_gate(
            demand_signal,
            context_label="el cálculo MRP",
            action_url=reverse("recetas:plan_produccion") + "#plan-pronosticos",
            action_label="Abrir pronóstico",
        )
        agregados: Dict[str, Dict[str, Any]] = {}
        unit_cost_cache: Dict[int, Decimal | None] = {}
        direct_base_replacement_cache: dict[int, list[dict[str, object]]] = {}
        lineas_sin_match = 0
        lineas_sin_cantidad = 0
        lineas_sin_costo = 0
        lineas_no_canonicas = 0
        lineas_base_directa = 0

        for l in receta.lineas.select_related("insumo", "insumo__proveedor_principal").all():
            _attach_linea_canonical_target(l)
            l.source_recipe = None
            l.source_code_kind = None
            l.source_active_presentaciones_count = 0
            l.uses_direct_base_in_final = False
            l.direct_base_replacement = None
            if l.insumo_id:
                l.source_code_kind = _derived_code_kind(l.insumo.codigo or "")
                source_recipe_id = _recipe_id_from_derived_code(l.insumo.codigo or "")
                if source_recipe_id:
                    l.source_recipe = (
                        Receta.objects.filter(pk=source_recipe_id).only("id", "nombre", "usa_presentaciones").first()
                    )
                    l.source_active_presentaciones_count = RecetaPresentacion.objects.filter(
                        receta_id=source_recipe_id,
                        activo=True,
                    ).count()
                    l.uses_direct_base_in_final = bool(
                        receta.tipo == Receta.TIPO_PRODUCTO_FINAL
                        and l.insumo.tipo_item == Insumo.TIPO_INTERNO
                        and l.source_code_kind == "PREPARACION"
                        and l.source_recipe
                        and l.source_recipe.usa_presentaciones
                        and l.source_active_presentaciones_count > 0
                    )
                    if l.uses_direct_base_in_final:
                        l.direct_base_replacement = _suggest_direct_base_replacement(
                            l,
                            cache=direct_base_replacement_cache,
                        )
            if not l.insumo_id:
                lineas_sin_match += 1
                key = f"(NO MATCH) {l.insumo_texto}"
                agregados.setdefault(
                    key,
                    {
                        "insumo": None,
                        "nombre": key,
                        "cantidad": Decimal("0"),
                        "unidad": l.unidad_texto,
                        "costo": Decimal("0"),
                        "origen": "Sin referencia",
                        "master_missing": [],
                        "stock_actual": Decimal("0"),
                        "faltante": Decimal("0"),
                        "alerta_capacidad": False,
                        "workflow_health_label": "Sin artículo estándar",
                        "workflow_health_tone": "danger",
                        "workflow_action_label": "Abrir centro de artículos",
                        "workflow_action_url": reverse("recetas:matching_pendientes") + f"?receta={receta.id}",
                        "workflow_action_method": "get",
                        "workflow_next": "Liga este componente al catálogo estándar.",
                    },
                )
                continue

            key = l.insumo.nombre
            if key not in agregados:
                article_class = _insumo_article_class(l.insumo)
                readiness = _insumo_erp_readiness(l.insumo)
                master_missing = list(readiness["missing"])
                if l.insumo.activo and not (l.insumo.codigo_point or "").strip():
                    master_missing.append("código comercial")
                agregados[key] = {
                    "insumo": l.insumo,
                    "nombre": key,
                    "cantidad": Decimal("0"),
                    "unidad": _linea_unit_code(l),
                    "costo": Decimal("0"),
                    "origen": article_class["label"],
                    "article_class_key": article_class["key"],
                    "article_class_label": article_class["label"],
                    "master_missing": master_missing,
                    "stock_actual": Decimal("0"),
                    "faltante": Decimal("0"),
                    "alerta_capacidad": False,
                    "workflow_health_label": "Lista para operar",
                    "workflow_health_tone": "success",
                    "workflow_action_label": "Ver artículo",
                    "workflow_action_url": reverse("maestros:insumo_update", args=[l.insumo_id]),
                    "workflow_action_method": "get",
                    "workflow_next": "Artículo listo para MRP rápido.",
                    "canonical_target": None,
                    "canonical_needs_repoint": False,
                    "uses_direct_base": False,
                    "direct_base_replacements": [],
                    "source_recipe_name": "",
                    "source_recipe_id": None,
                }

            if getattr(l, "canonical_needs_repoint", False) and getattr(l, "canonical_target", None):
                lineas_no_canonicas += 1
                agregados[key]["canonical_needs_repoint"] = True
                agregados[key]["canonical_target"] = l.canonical_target
                if agregados[key]["workflow_health_label"] == "Lista para operar":
                    agregados[key]["workflow_health_label"] = "Fuera de estándar"
                    agregados[key]["workflow_health_tone"] = "warning"
                    agregados[key]["workflow_action_label"] = "Usar artículo estándar"
                    agregados[key]["workflow_action_url"] = reverse("recetas:receta_detail", args=[receta.id])
                    agregados[key]["workflow_action_method"] = "get"
                    agregados[key]["workflow_next"] = "Reapunta el componente al artículo estándar antes de usarlo en compras o MRP."

            if getattr(l, "uses_direct_base_in_final", False):
                lineas_base_directa += 1
                agregados[key]["uses_direct_base"] = True
                if getattr(l, "source_recipe", None):
                    agregados[key]["source_recipe_name"] = getattr(l.source_recipe, "nombre", "") or ""
                    agregados[key]["source_recipe_id"] = getattr(l.source_recipe, "id", None)
                if getattr(l, "direct_base_replacement", None):
                    suggestion = l.direct_base_replacement
                    suggestion_name = suggestion["insumo"].nombre
                    existing_names = {
                        item["insumo"].nombre
                        for item in agregados[key]["direct_base_replacements"]
                    }
                    if suggestion_name not in existing_names:
                        agregados[key]["direct_base_replacements"].append(suggestion)
                if agregados[key]["workflow_health_label"] in {"Lista para operar", "Fuera de estándar"}:
                    agregados[key]["workflow_health_label"] = "Usa base sin presentación"
                    agregados[key]["workflow_health_tone"] = "danger"
                    if agregados[key]["direct_base_replacements"]:
                        agregados[key]["workflow_action_label"] = "Aplicar sugeridos"
                        agregados[key]["workflow_action_url"] = reverse("recetas:receta_apply_direct_base_replacements", args=[receta.id])
                        agregados[key]["workflow_action_method"] = "post"
                        agregados[key]["workflow_next"] = "Sustituye la base directa por la presentación derivada sugerida para estabilizar costeo y MRP."
                    else:
                        agregados[key]["workflow_action_label"] = "Corregir receta"
                        agregados[key]["workflow_action_url"] = reverse("recetas:receta_detail", args=[receta.id])
                        agregados[key]["workflow_action_method"] = "get"
                        agregados[key]["workflow_next"] = "Sustituye la base directa por la presentación derivada correcta para estabilizar costeo y MRP."

            qty_base = Decimal(str(l.cantidad or 0))
            if qty_base <= 0:
                lineas_sin_cantidad += 1
                agregados[key]["workflow_health_label"] = "Sin cantidad"
                agregados[key]["workflow_health_tone"] = "warning"
                agregados[key]["workflow_action_label"] = "Revisar receta"
                agregados[key]["workflow_action_url"] = reverse("recetas:receta_detail", args=[receta.id])
                agregados[key]["workflow_action_method"] = "get"
                agregados[key]["workflow_next"] = "Completa cantidad en la línea BOM para usar este componente."
                continue

            qty = qty_base * multiplicador
            agregados[key]["cantidad"] += qty

            unit_cost = _linea_unit_cost(l, unit_cost_cache)
            if unit_cost is None or unit_cost <= 0:
                lineas_sin_costo += 1
                unit_cost = Decimal("0")
                agregados[key]["workflow_health_label"] = "Sin costo"
                agregados[key]["workflow_health_tone"] = "warning"
                agregados[key]["workflow_action_label"] = "Completar costo"
                agregados[key]["workflow_action_url"] = reverse("maestros:insumo_list") + "?costo_status=sin_costo"
                agregados[key]["workflow_action_method"] = "get"
                agregados[key]["workflow_next"] = "Asigna costo vigente al artículo antes de usarlo en compras."

            agregados[key]["costo"] += qty * unit_cost

        derived_parent_row = _build_derived_parent_requirement(receta, multiplicador)
        if derived_parent_row:
            agregados[derived_parent_row["key"]] = {
                **derived_parent_row,
                "canonical_target": None,
                "canonical_needs_repoint": False,
                "uses_direct_base": False,
                "direct_base_replacements": [],
            }

        insumo_ids = [item["insumo"].id for item in agregados.values() if item.get("insumo")]
        existencias_map = {
            e.insumo_id: Decimal(str(e.stock_actual or 0))
            for e in ExistenciaInsumo.objects.filter(insumo_id__in=insumo_ids)
        }
        master_incompletos = 0
        alertas_capacidad = 0
        for row in agregados.values():
            insumo_obj = row.get("insumo")
            if row.get("is_derived_parent"):
                if row["alerta_capacidad"]:
                    alertas_capacidad += 1
                    row["workflow_health_label"] = "Preparar padre"
                    row["workflow_health_tone"] = "warning"
                    row["workflow_action_label"] = "Abrir padre"
                    row["workflow_action_url"] = row["detail_url"]
                    row["workflow_action_method"] = "get"
                    row["workflow_next"] = "Programa el producto padre para cubrir la presentación derivada."
                if Decimal(str(row["costo_unitario"] or 0)) <= 0:
                    lineas_sin_costo += 1
                    row["workflow_health_label"] = "Sin costo base padre"
                    row["workflow_health_tone"] = "warning"
                    row["workflow_action_label"] = "Abrir padre"
                    row["workflow_action_url"] = row["detail_url"]
                    row["workflow_action_method"] = "get"
                    row["workflow_next"] = "Cierra el costeo del producto padre antes de usar esta presentación."
                continue

            if not insumo_obj:
                continue
            row["stock_actual"] = existencias_map.get(insumo_obj.id, Decimal("0"))
            faltante = Decimal(str(row["cantidad"] or 0)) - Decimal(str(row["stock_actual"] or 0))
            row["faltante"] = faltante if faltante > 0 else Decimal("0")
            row["alerta_capacidad"] = row["faltante"] > 0
            if row["alerta_capacidad"]:
                alertas_capacidad += 1
                if row["origen"] == "Insumo interno":
                    row["workflow_health_label"] = "Producir interno"
                    row["workflow_health_tone"] = "warning"
                    row["workflow_action_label"] = "Revisar base"
                    row["workflow_action_url"] = reverse("recetas:recetas_list") + f"?q={insumo_obj.nombre}"
                    row["workflow_action_method"] = "get"
                    row["workflow_next"] = "Programa producción interna para cubrir el faltante."
                else:
                    row["workflow_health_label"] = "Comprar"
                    row["workflow_health_tone"] = "warning"
                    row["workflow_action_label"] = "Ir a compras"
                    row["workflow_action_url"] = reverse("compras:solicitudes")
                    row["workflow_action_method"] = "get"
                    row["workflow_next"] = "Genera solicitud u orden para cubrir el faltante."
            if row["master_missing"]:
                master_incompletos += 1
                if row["workflow_health_label"] == "Lista para operar":
                    row["workflow_health_label"] = "Maestro incompleto"
                    row["workflow_health_tone"] = "warning"
                    row["workflow_action_label"] = "Abrir maestro"
                    row["workflow_action_url"] = reverse("maestros:insumo_update", args=[insumo_obj.id])
                    row["workflow_action_method"] = "get"
                    row["workflow_next"] = "Completa el maestro para usar el artículo sin ambigüedad."

        master_demand_critical = bool(
            master_incompletos
            and demand_signal
            and demand_signal.get("historico_tone") in {"success", "warning"}
        )
        items_sorted = sorted(agregados.values(), key=lambda x: x["nombre"])
        lineas_base_directa_sugeridas = sum(
            1 for item in items_sorted if item.get("direct_base_replacements")
        )
        chain_cards: list[dict[str, Any]] = []
        chain_detail_rows: list[dict[str, Any]] = []
        quality_cards: list[dict[str, Any]] = []
        master_blocker_groups: dict[str, dict[str, Any]] = {}
        master_blocker_missing_groups: dict[str, dict[str, Any]] = {}
        master_blocker_detail_rows: list[dict[str, Any]] = []
        if lineas_base_directa:
            chain_cards.append(
                {
                    "key": "base_directa",
                    "label": "Bases sin presentación detectadas",
                    "count": lineas_base_directa,
                    "tone": "danger",
                    "detail": (
                        f"{lineas_base_directa_sugeridas} componente(s) ya tienen presentación sugerida."
                        if lineas_base_directa_sugeridas
                        else "Todavía requieren ajuste manual en la receta."
                    ),
                }
            )
        if lineas_no_canonicas:
            chain_cards.append(
                {
                    "key": "no_canonico",
                    "label": "Artículos fuera de maestro",
                    "count": lineas_no_canonicas,
                    "tone": "warning",
                    "detail": "Conviene ajustar la estructura al artículo estándar antes de comprar o producir.",
                }
            )
        if master_incompletos:
            maestro_demanda_tone = "danger" if master_demand_critical else "warning"
            chain_cards.append(
                {
                    "key": "maestro_bloqueando",
                    "label": (
                        "Demanda crítica bloqueada por maestro"
                        if maestro_demanda_tone == "danger"
                        else "Maestro bloqueando MRP"
                    ),
                    "count": master_incompletos,
                    "tone": maestro_demanda_tone,
                    "detail": (
                        "La receta tiene base comercial suficiente, pero sigue frenada por artículos incompletos en maestro."
                        if maestro_demanda_tone == "danger"
                        else "Hay artículos ligados sin datos suficientes en maestro para operar con estabilidad."
                    ),
                }
            )
        if lineas_sin_match:
            quality_cards.append({"key": "sin_match", "label": "Sin artículo estándar", "count": lineas_sin_match, "tone": "danger", "action_label": "Resolver catálogo", "action_url": reverse("recetas:matching_pendientes") + f"?receta={receta.id}"})
        if lineas_sin_cantidad:
            quality_cards.append({"key": "sin_cantidad", "label": "Sin cantidad", "count": lineas_sin_cantidad, "tone": "warning", "action_label": "Revisar receta", "action_url": reverse("recetas:receta_detail", args=[receta.id])})
        if lineas_sin_costo:
            quality_cards.append({"key": "sin_costo", "label": "Sin costo", "count": lineas_sin_costo, "tone": "warning", "action_label": "Completar costos", "action_url": reverse("maestros:insumo_list") + "?costo_status=sin_costo"})
        if lineas_base_directa:
            quality_cards.append(
                {
                    "key": "base_directa",
                    "label": "Usa base sin presentación",
                    "count": lineas_base_directa,
                    "tone": "danger",
                    "action_label": (
                        "Aplicar derivados sugeridos"
                        if lineas_base_directa_sugeridas
                        else "Corregir receta"
                    ),
                    "action_url": (
                        reverse("recetas:receta_apply_direct_base_replacements", args=[receta.id])
                        if lineas_base_directa_sugeridas
                        else reverse("recetas:receta_detail", args=[receta.id])
                    ),
                    "action_method": "post" if lineas_base_directa_sugeridas else "get",
                    "detail": (
                        f"{lineas_base_directa_sugeridas} línea(s) ya tienen derivado sugerido."
                        if lineas_base_directa_sugeridas
                        else "Revisa la receta y sustituye la base directa."
                    ),
                }
            )
        if lineas_no_canonicas:
            quality_cards.append({"key": "no_canonico", "label": "Fuera de estándar", "count": lineas_no_canonicas, "tone": "warning", "action_label": "Usar artículo estándar", "action_url": reverse("recetas:receta_detail", args=[receta.id])})
        if alertas_capacidad:
            quality_cards.append({"key": "stock_insuficiente", "label": "Stock insuficiente", "count": alertas_capacidad, "tone": "danger", "action_label": "Ver existencias", "action_url": reverse("inventario:existencias")})
        if master_incompletos:
            maestro_demanda_tone = "danger" if master_demand_critical else "warning"
            quality_cards.append(
                {
                    "key": "maestro_incompleto",
                    "label": (
                        "Demanda crítica bloqueada por maestro"
                        if maestro_demanda_tone == "danger"
                        else "Maestro incompleto"
                    ),
                    "count": master_incompletos,
                    "tone": maestro_demanda_tone,
                    "action_label": "Abrir maestro",
                    "action_url": reverse("maestros:insumo_list") + "?enterprise_status=incompletos&usage_scope=recipes",
                    "detail": (
                        "La señal comercial ya es utilizable, pero MRP sigue bloqueado por artículos incompletos en maestro."
                        if maestro_demanda_tone == "danger"
                        else "Cierra los artículos incompletos antes de usar el cálculo de forma estable."
                    ),
                }
            )
        if commercial_gate["tone"] != "success":
            quality_cards.append(
                {
                    "key": "demanda_fragil",
                    "label": "Base comercial en revisión" if commercial_gate["tone"] == "warning" else "Base comercial frágil",
                    "count": commercial_gate["blockers"],
                    "tone": commercial_gate["tone"],
                    "action_label": commercial_gate["action_label"],
                    "action_url": commercial_gate["action_url"],
                    "detail": commercial_gate["detail"],
                }
            )

        for row in items_sorted:
            if row.get("uses_direct_base"):
                replacement = (row.get("direct_base_replacements") or [None])[0]
                replacement_name = replacement["insumo"].nombre if replacement else "Sin derivado sugerido"
                replacement_reason = replacement.get("reason") if replacement else "Revisa la base y define la presentación derivada correcta."
                replacement_qty = replacement.get("replacement_quantity") if replacement else None
                replacement_unit = getattr(getattr(replacement.get("insumo"), "unidad_base", None), "codigo", "") if replacement else ""
                chain_detail_rows.append(
                    {
                        "type": "Base sin presentación",
                        "name": row["nombre"],
                        "source_recipe_name": row.get("source_recipe_name") or "Base sin origen",
                        "detail": (
                            f"Derivado sugerido: {replacement_name}"
                            + (
                                f" · {Decimal(str(replacement_qty)):.2f} {replacement_unit}"
                                if replacement_qty is not None and replacement_unit
                                else ""
                            )
                        ),
                        "note": replacement_reason,
                        "action_label": "Aplicar derivados sugeridos" if replacement else "Revisar receta",
                        "action_url": (
                            reverse("recetas:receta_apply_direct_base_replacements", args=[receta.id])
                            if replacement
                            else reverse("recetas:receta_detail", args=[receta.id])
                        ),
                        "action_method": "post" if replacement else "get",
                    }
                )
            elif row.get("canonical_needs_repoint") and row.get("canonical_target"):
                chain_detail_rows.append(
                    {
                        "type": "Fuera de estándar",
                        "name": row["nombre"],
                        "source_recipe_name": row.get("source_recipe_name") or "Catálogo",
                        "detail": f"Artículo estándar sugerido: {row['canonical_target'].nombre}",
                        "note": "Normaliza la línea para evitar costos o compras divididas por variante.",
                        "action_label": "Usar artículo estándar",
                        "action_url": reverse("recetas:receta_detail", args=[receta.id]),
                        "action_method": "get",
                    }
                )
            if not row.get("master_missing"):
                continue
            class_key = str(row.get("article_class_key") or Insumo.TIPO_MATERIA_PRIMA)
            class_label = str(row.get("article_class_label") or row.get("origen") or "Artículo")
            group = master_blocker_groups.setdefault(
                class_key,
                {
                    "class_key": class_key,
                    "class_label": class_label,
                    "count": 0,
                    "missing_totals": defaultdict(int),
                },
            )
            group["count"] += 1
            for missing_label in row.get("master_missing") or []:
                group["missing_totals"][missing_label] += 1
                missing_key = _missing_field_to_filter_key(missing_label) or "maestro"
                missing_group = master_blocker_missing_groups.setdefault(
                    missing_key,
                    {
                        "missing_key": missing_key,
                        "missing_label": missing_label,
                        "count": 0,
                        "class_totals": defaultdict(int),
                    },
                )
                missing_group["count"] += 1
                missing_group["class_totals"][class_label] += 1

            primary_missing = (row.get("master_missing") or [None])[0]
            action_meta = _enterprise_blocker_action_meta_for_recipes(
                row["nombre"],
                class_key,
                primary_missing,
                insumo_id=getattr(row.get("insumo"), "id", None) or row.get("insumo_id"),
                usage_scope="recipes",
            )
            master_blocker_detail_rows.append(
                {
                    "class_key": class_key,
                    "class_label": class_label,
                    "name": row["nombre"],
                    "missing_field": _missing_field_to_filter_key(primary_missing) or "maestro",
                    "missing": ", ".join(row.get("master_missing") or []),
                    "detail": "Completa el maestro para liberar costeo, MRP y compras del cálculo rápido.",
                    "action_label": action_meta["label"],
                    "action_detail": action_meta["detail"],
                    "action_url": action_meta["url"],
                    "edit_url": action_meta["edit_url"],
                    "tone": "warning",
                }
            )
            chain_detail_rows.append(
                {
                    "type": "Maestro incompleto",
                    "name": row["nombre"],
                    "source_recipe_name": class_label,
                    "detail": ", ".join(row.get("master_missing") or []),
                    "note": "Completa el maestro para liberar costeo, MRP y compras del cálculo rápido.",
                    "action_label": action_meta["label"],
                    "action_url": action_meta["url"],
                    "action_method": "get",
                }
            )

        master_blocker_class_cards: list[dict[str, Any]] = []
        for group in sorted(master_blocker_groups.values(), key=lambda item: (-item["count"], item["class_label"])):
            dominant_label = ""
            dominant_count = 0
            for missing_label, count in dict(group["missing_totals"]).items():
                if count > dominant_count:
                    dominant_label = missing_label
                    dominant_count = count
            action_label, action_detail = _enterprise_blocker_label_detail_for_missing_recipes(dominant_label)
            query = {
                "tipo_item": group["class_key"],
                "enterprise_status": "incompletos",
                "usage_scope": "recipes",
            }
            filter_key = _missing_field_to_filter_key(dominant_label)
            if filter_key:
                query["missing_field"] = filter_key
            master_blocker_class_cards.append(
                {
                    "key": group["class_key"],
                    "class_label": group["class_label"],
                    "count": group["count"],
                    "dominant_label": dominant_label or "maestro incompleto",
                    "dominant_count": dominant_count,
                    "action_label": action_label,
                    "action_detail": action_detail,
                    "action_url": reverse("maestros:insumo_list") + f"?{urlencode(query)}",
                }
            )
        master_blocker_missing_cards: list[dict[str, Any]] = []
        for missing_group in sorted(
            master_blocker_missing_groups.values(),
            key=lambda item: (-item["count"], item["missing_label"]),
        ):
            dominant_class_label = ""
            dominant_class_count = 0
            for class_label, count in dict(missing_group["class_totals"]).items():
                if count > dominant_class_count:
                    dominant_class_label = class_label
                    dominant_class_count = count
            action_label, action_detail = _enterprise_blocker_label_detail_for_missing_recipes(
                missing_group["missing_label"]
            )
            query = {
                "enterprise_status": "incompletos",
                "usage_scope": "recipes",
                "missing_field": missing_group["missing_key"],
            }
            master_blocker_missing_cards.append(
                {
                    "key": missing_group["missing_key"],
                    "missing_label": missing_group["missing_label"],
                    "count": missing_group["count"],
                    "dominant_class_label": dominant_class_label or "Artículo",
                    "dominant_class_count": dominant_class_count,
                    "action_label": action_label,
                    "action_detail": action_detail,
                    "action_url": reverse("maestros:insumo_list") + f"?{urlencode(query)}",
                }
            )

        if lineas_sin_match or lineas_sin_cantidad or lineas_sin_costo or lineas_base_directa:
            health_label = "Con bloqueos operativos"
            health_tone = "danger"
            health_detail = "El cálculo rápido tiene bloqueos de datos que deben corregirse antes de usarlo para compras o costeo."
        elif alertas_capacidad or master_incompletos or lineas_no_canonicas:
            health_label = "Atención operativa"
            health_tone = "warning"
            health_detail = "El cálculo está listo, pero hay faltantes de stock o artículos incompletos en el maestro."
        else:
            health_label = "Lista para operar"
            health_tone = "success"
            health_detail = "El cálculo no tiene bloqueos críticos de datos ni faltantes visibles."

        downstream_handoff_rows = [
            {
                "label": "MRP",
                "owner": "Planeación / Producción",
                "status": "Listo" if health_tone == "success" else "Bloqueado",
                "tone": "success" if health_tone == "success" else ("warning" if health_tone == "warning" else "danger"),
                "blockers": lineas_sin_match + lineas_sin_cantidad + lineas_sin_costo + lineas_base_directa + master_incompletos + int(commercial_gate["blockers"]),
                "completion": 100 if health_tone == "success" and commercial_gate["tone"] == "success" else 70 if health_tone == "warning" or commercial_gate["tone"] == "warning" else 35,
                "depends_on": "Catálogo maestro + estructura BOM + señal comercial",
                "exit_criteria": "La simulación debe quedar sin referencias abiertas, sin base directa, con costo operativo válido y con base comercial utilizable.",
                "detail": (
                    "El cálculo ya puede alimentar planeación y explosión de demanda."
                    if health_tone == "success" and commercial_gate["tone"] == "success"
                    else commercial_gate["detail"] if commercial_gate["tone"] != "success" else health_detail
                ),
                "next_step": (
                    "Abrir plan"
                    if health_tone == "success" and commercial_gate["tone"] == "success"
                    else commercial_gate["next_step"] if commercial_gate["tone"] != "success" else "Cerrar bloqueos del cálculo"
                ),
                "url": reverse("recetas:plan_produccion") if health_tone == "success" and commercial_gate["tone"] == "success" else commercial_gate["action_url"] if commercial_gate["tone"] != "success" else reverse("recetas:receta_detail", args=[receta.id]),
                "cta": "Abrir plan" if health_tone == "success" and commercial_gate["tone"] == "success" else commercial_gate["action_label"] if commercial_gate["tone"] != "success" else "Revisar receta",
            },
            {
                "label": "Compras",
                "owner": "Compras / Planeación",
                "status": "Listo" if health_tone == "success" and alertas_capacidad > 0 else "Bloqueado",
                "tone": "success" if health_tone == "success" and alertas_capacidad > 0 else "warning",
                "blockers": lineas_sin_match + lineas_sin_cantidad + lineas_sin_costo + lineas_base_directa + master_incompletos,
                "completion": 100 if health_tone == "success" and alertas_capacidad > 0 else 55,
                "depends_on": "Stock faltante + maestro cerrado",
                "exit_criteria": "Compras solo debe recibir faltantes ya validados y con artículos completos en maestro.",
                "detail": (
                    "El cálculo ya detectó faltantes válidos para abastecimiento documental."
                    if health_tone == "success" and alertas_capacidad > 0
                    else "Compras sigue en espera hasta cerrar la simulación o detectar faltantes reales de stock."
                ),
                "next_step": "Abrir compras" if health_tone == "success" and alertas_capacidad > 0 else "Validar simulación",
                "url": reverse("compras:solicitudes") if health_tone == "success" and alertas_capacidad > 0 else reverse("recetas:receta_detail", args=[receta.id]),
                "cta": "Abrir compras" if health_tone == "success" and alertas_capacidad > 0 else "Revisar receta",
            },
            {
                "label": "Inventario",
                "owner": "Inventario / Almacén",
                "status": "Listo" if lineas_sin_match == 0 and master_incompletos == 0 else "Bloqueado",
                "tone": "success" if lineas_sin_match == 0 and master_incompletos == 0 else "warning",
                "blockers": lineas_sin_match + master_incompletos,
                "completion": 100 if lineas_sin_match == 0 and master_incompletos == 0 else 60,
                "depends_on": "Artículo maestro + existencia actual",
                "exit_criteria": "Inventario debe leer artículos maestros y existencias operativas válidas.",
                "detail": (
                    "Inventario ya puede usar la simulación con artículos maestros consistentes."
                    if lineas_sin_match == 0 and master_incompletos == 0
                    else "Inventario todavía requiere cerrar referencias maestras o artículos incompletos."
                ),
                "next_step": "Abrir inventario" if lineas_sin_match == 0 and master_incompletos == 0 else "Cerrar maestro",
                "url": reverse("inventario:existencias") if lineas_sin_match == 0 and master_incompletos == 0 else reverse("maestros:insumo_list"),
                "cta": "Abrir inventario" if lineas_sin_match == 0 and master_incompletos == 0 else "Abrir maestro",
            },
        ]

        master_focus_rows = list(master_blocker_detail_rows[:3])
        if master_focus_rows:
            first_master_focus = master_focus_rows[0]
            master_focus = {
                **first_master_focus,
                "label": f"{first_master_focus['class_label']} · {first_master_focus['missing_field']}",
                "summary": (
                    f"El cálculo MRP sigue condicionado por {first_master_focus['name']} "
                    f"({first_master_focus['missing_field']})."
                ),
                "tone": "danger" if first_master_focus.get("tone") == "danger" else "warning",
            }
        else:
            master_focus = {
                "class_label": "Maestro",
                "missing_field": "sin bloqueos",
            "label": "Maestro ERP al día",
                "summary": "No hay artículos bloqueando el cálculo rápido por faltantes del maestro.",
                "action_label": "Abrir maestro",
                "action_detail": "Puedes revisar el catálogo general para seguimiento preventivo.",
                "action_url": reverse("maestros:insumo_list"),
                "tone": "success",
                "count": 0,
            }

        workflow_rows = [
            {
                "step": "01",
                "title": "Estructura BOM",
                "owner": "Producción / Costeo",
                "open": lineas_sin_cantidad + lineas_base_directa,
                "closed": max(len(items_sorted) - (lineas_sin_cantidad + lineas_base_directa), 0),
                "completion": 100 if (lineas_sin_cantidad + lineas_base_directa) == 0 else max(20, 100 - ((lineas_sin_cantidad + lineas_base_directa) * 15)),
                "tone": "success" if (lineas_sin_cantidad + lineas_base_directa) == 0 else "warning",
                "detail": (
                    "La estructura está lista para operar."
                    if (lineas_sin_cantidad + lineas_base_directa) == 0
                    else f"{lineas_sin_cantidad} sin cantidad y {lineas_base_directa} con base directa."
                ),
                "next_step": (
                    "Mantener estructura estable."
                    if (lineas_sin_cantidad + lineas_base_directa) == 0
                    else "Completar cantidades y sustituir bases directas por derivados."
                ),
                "action_label": "Abrir receta",
                "action_href": reverse("recetas:receta_detail", args=[receta.id]),
            },
            {
                "step": "02",
                "title": "Maestro del artículo",
                "owner": "Maestros / DG",
                "open": lineas_sin_match + master_incompletos + lineas_no_canonicas,
                "closed": max(len(items_sorted) - (lineas_sin_match + master_incompletos + lineas_no_canonicas), 0),
                "completion": (
                    12
                    if master_demand_critical
                    else 100
                    if (lineas_sin_match + master_incompletos + lineas_no_canonicas) == 0
                    else max(20, 100 - ((lineas_sin_match + master_incompletos + lineas_no_canonicas) * 12))
                ),
                "tone": (
                    "danger"
                    if master_demand_critical
                    else "success"
                    if (lineas_sin_match + master_incompletos + lineas_no_canonicas) == 0
                    else ("danger" if lineas_sin_match else "warning")
                ),
                "detail": (
                    "La receta tiene demanda utilizable, pero sigue bloqueada por artículos incompletos del maestro."
                    if master_demand_critical
                    else "Todos los componentes usan artículo maestro estable."
                    if (lineas_sin_match + master_incompletos + lineas_no_canonicas) == 0
                    else f"{lineas_sin_match} sin alta, {master_incompletos} incompletos y {lineas_no_canonicas} fuera de estándar."
                ),
                "next_step": (
                    "Cerrar primero el artículo maestro crítico antes de liberar compras o producción."
                    if master_demand_critical
                    else "Mantener maestro preventivo."
                    if (lineas_sin_match + master_incompletos + lineas_no_canonicas) == 0
                    else "Cerrar altas maestras, completar datos y normalizar variantes."
                ),
                "action_label": "Abrir maestro",
                "action_href": reverse("maestros:insumo_list"),
            },
            {
                "step": "03",
                "title": "Capacidad operativa",
                "owner": "Inventario / Compras",
                "open": alertas_capacidad + lineas_sin_costo,
                "closed": max(len(items_sorted) - (alertas_capacidad + lineas_sin_costo), 0),
                "completion": 100 if (alertas_capacidad + lineas_sin_costo) == 0 else max(20, 100 - ((alertas_capacidad + lineas_sin_costo) * 15)),
                "tone": "success" if (alertas_capacidad + lineas_sin_costo) == 0 else ("danger" if alertas_capacidad else "warning"),
                "detail": (
                    "Stock y costo listos para operar."
                    if (alertas_capacidad + lineas_sin_costo) == 0
                    else f"{alertas_capacidad} faltantes y {lineas_sin_costo} artículos sin costo."
                ),
                "next_step": (
                    "Usar resultado en plan y compras."
                    if (alertas_capacidad + lineas_sin_costo) == 0
                    else "Cubrir faltantes o completar costos antes de liberar abastecimiento."
                ),
                "action_label": "Abrir existencias",
                "action_href": reverse("inventario:existencias"),
            },
        ]

        if health_tone == "danger":
            erp_command_center = {
                "owner": "Producción / Costeo",
                "status": "Crítico",
                "tone": "danger",
                "blockers": lineas_sin_match + lineas_sin_cantidad + lineas_sin_costo + lineas_base_directa + alertas_capacidad,
                "next_step": "Corregir estructura BOM y bloqueos críticos antes de usar este cálculo para compras o producción.",
                "url": reverse("recetas:receta_detail", args=[receta.id]),
                "cta": "Abrir receta operativa",
            }
        elif master_demand_critical:
            erp_command_center = {
                "owner": "Maestros / DG",
                "status": "Crítico",
                "tone": "danger",
                "blockers": master_incompletos,
                "next_step": (
                    "Cierra primero el artículo maestro crítico antes de usar esta receta para MRP, compras o producción."
                ),
                "url": reverse("maestros:insumo_list") + "?enterprise_status=incompletos&usage_scope=recipes",
                "cta": "Cerrar maestro crítico",
            }
        elif commercial_gate["tone"] == "danger":
            erp_command_center = {
                "owner": "Ventas / Planeación",
                "status": "Frágil",
                "tone": "danger",
                "blockers": int(commercial_gate["blockers"]),
                "next_step": commercial_gate["next_step"],
                "url": commercial_gate["action_url"],
                "cta": commercial_gate["action_label"],
            }
        elif commercial_gate["tone"] == "warning":
            erp_command_center = {
                "owner": "Ventas / Planeación",
                "status": "En revisión",
                "tone": "warning",
                "blockers": int(commercial_gate["blockers"]),
                "next_step": commercial_gate["next_step"],
                "url": commercial_gate["action_url"],
                "cta": commercial_gate["action_label"],
            }
        elif master_incompletos or lineas_no_canonicas:
            erp_command_center = {
                "owner": "Maestros / DG",
                "status": "En revisión",
                "tone": "warning",
                "blockers": master_incompletos + lineas_no_canonicas,
                "next_step": "Cerrar maestro del artículo y normalizar variantes antes de liberar operación completa.",
                "url": reverse("maestros:insumo_list") + "?enterprise_status=incompletos&usage_scope=recipes",
                "cta": "Abrir maestro ERP",
            }
        else:
            erp_command_center = {
                "owner": "Inventario / Compras",
                "status": "Estable",
                "tone": "success",
                "blockers": 0,
                "next_step": "El cálculo está listo para abastecimiento, costeo y validación documental.",
                "url": reverse("recetas:plan_produccion"),
                "cta": "Abrir plan operativo",
            }

        upstream_dependency_rows: list[dict[str, Any]] = []
        if demand_signal:
            commercial_blockers = 0
            if demand_signal["historico_tone"] == "danger":
                commercial_blockers += 1
            if demand_signal["forecast_tone"] in {"danger", "warning"}:
                commercial_blockers += 1
            if demand_signal["alignment_tone"] == "danger":
                commercial_blockers += 1
            if commercial_blockers == 0:
                commercial_status = "Lista"
                commercial_tone = "success"
                commercial_completion = 100
                commercial_next = "Usar esta señal como referencia para plan y abastecimiento."
            elif commercial_blockers == 1:
                commercial_status = "Utilizable"
                commercial_tone = "warning"
                commercial_completion = 70
                commercial_next = "Revisar la brecha comercial antes de comprometer compras."
            else:
                commercial_status = "Frágil"
                commercial_tone = "danger"
                commercial_completion = 35
                commercial_next = "Reforzar histórico/forecast antes de usar esta receta como base de abastecimiento."
            upstream_dependency_rows.append(
                {
                    "label": "Demanda comercial",
                    "owner": "Ventas / Planeación",
                    "status": commercial_status,
                    "tone": commercial_tone,
                    "blockers": commercial_blockers,
                    "completion": commercial_completion,
                    "depends_on": "Ventas históricas + forecast + solicitud",
                    "exit_criteria": "La receta debe tener base histórica utilizable y una señal comercial coherente.",
                    "detail": demand_signal["historico_detail"],
                    "next_step": commercial_next,
                    "url": reverse("recetas:plan_produccion"),
                    "cta": "Abrir plan",
                }
            )
        if master_demand_critical:
            upstream_dependency_rows.append(
                {
                    "label": "Maestro crítico por demanda",
                    "owner": "Maestros / DG",
                    "status": "Crítico",
                    "tone": "danger",
                    "blockers": master_incompletos,
                    "completion": 12,
                    "depends_on": "Artículo maestro + señal comercial suficiente",
                    "exit_criteria": "Los artículos críticos del BOM deben quedar completos en maestro antes de liberar abastecimiento.",
                    "detail": master_focus["summary"],
                    "next_step": "Cerrar primero el artículo maestro crítico antes de confiar en este cálculo.",
                    "url": reverse("maestros:insumo_list") + "?enterprise_status=incompletos&usage_scope=recipes&impact_scope=critical",
                    "cta": "Cerrar prioridad crítica",
                }
            )

        allowed_focus_kinds = {"quality", "master", "master_missing", "chain"}
        if focus_kind not in allowed_focus_kinds:
            focus_kind = ""
        selected_focus_kind = focus_kind
        selected_focus_key = focus_key if focus_kind else ""
        focus_summary = ""

        filtered_items = items_sorted
        filtered_master_rows = master_blocker_detail_rows[:12]
        filtered_chain_rows = chain_detail_rows[:12]

        if selected_focus_kind == "quality" and selected_focus_key:
            if selected_focus_key == "sin_match":
                filtered_items = [item for item in items_sorted if not item.get("insumo")]
                focus_summary = "Filtrando solo componentes pendientes de catálogo."
            elif selected_focus_key == "sin_cantidad":
                filtered_items = [item for item in items_sorted if item.get("workflow_health_label") == "Sin cantidad"]
                focus_summary = "Filtrando líneas BOM sin cantidad."
            elif selected_focus_key == "sin_costo":
                filtered_items = [item for item in items_sorted if item.get("workflow_health_label") == "Sin costo"]
                focus_summary = "Filtrando artículos sin costo vigente."
            elif selected_focus_key == "base_directa":
                filtered_items = [item for item in items_sorted if item.get("uses_direct_base")]
                filtered_chain_rows = [row for row in chain_detail_rows if row.get("type") == "Base sin presentación"][:12]
                focus_summary = "Filtrando bases completas que deben sustituirse por presentaciones derivadas."
            elif selected_focus_key == "no_canonico":
                filtered_items = [item for item in items_sorted if item.get("canonical_needs_repoint")]
                filtered_chain_rows = [row for row in chain_detail_rows if row.get("type") == "Fuera de estándar"][:12]
                focus_summary = "Filtrando artículos fuera de maestro en la explosión MRP."
            elif selected_focus_key == "stock_insuficiente":
                filtered_items = [item for item in items_sorted if item.get("alerta_capacidad")]
                focus_summary = "Filtrando faltantes de stock para compra o producción."
            elif selected_focus_key == "maestro_incompleto":
                filtered_items = [item for item in items_sorted if item.get("master_missing")]
                filtered_master_rows = [row for row in master_blocker_detail_rows if row.get("missing")] [:12]
                focus_summary = "Filtrando artículos bloqueados por maestro incompleto."
        elif selected_focus_kind == "master" and selected_focus_key:
            filtered_items = [item for item in items_sorted if item.get("master_missing") and str(item.get("article_class_key") or Insumo.TIPO_MATERIA_PRIMA) == selected_focus_key]
            filtered_master_rows = [row for row in master_blocker_detail_rows if row.get("class_key") == selected_focus_key][:12]
            focus_summary = "Filtrando solo la clase del maestro seleccionada."
        elif selected_focus_kind == "master_missing" and selected_focus_key:
            filtered_items = [
                item
                for item in items_sorted
                if item.get("master_missing")
                and selected_focus_key in {
                    _missing_field_to_filter_key(missing_label) or "maestro"
                    for missing_label in (item.get("master_missing") or [])
                }
            ]
            filtered_master_rows = [
                row for row in master_blocker_detail_rows if row.get("missing_field") == selected_focus_key
            ][:12]
            focus_summary = "Filtrando solo el dato faltante del maestro seleccionado."
        elif selected_focus_kind == "chain" and selected_focus_key:
            if selected_focus_key == "base_directa":
                filtered_items = [item for item in items_sorted if item.get("uses_direct_base")]
                filtered_chain_rows = [row for row in chain_detail_rows if row.get("type") == "Base sin presentación"][:12]
                focus_summary = "Filtrando bloqueos por uso de base sin presentación."
            elif selected_focus_key == "no_canonico":
                filtered_items = [item for item in items_sorted if item.get("canonical_needs_repoint")]
                filtered_chain_rows = [row for row in chain_detail_rows if row.get("type") == "Fuera de estándar"][:12]
                focus_summary = "Filtrando bloqueos de canonicidad en la cadena."
            elif selected_focus_key == "maestro_bloqueando":
                filtered_items = [item for item in items_sorted if item.get("master_missing")]
                filtered_chain_rows = [row for row in chain_detail_rows if row.get("type") == "Maestro incompleto"][:12]
                focus_summary = "Filtrando artículos que bloquean la cadena por maestro incompleto."

        resultado = {
            "receta": receta,
            "multiplicador": multiplicador,
            "items": filtered_items,
            "all_items": items_sorted,
            "costo_total": sum((Decimal(str(i["costo"] or 0)) for i in items_sorted), Decimal("0")),
            "lineas_sin_match": lineas_sin_match,
            "lineas_sin_cantidad": lineas_sin_cantidad,
            "lineas_sin_costo": lineas_sin_costo,
            "lineas_no_canonicas": lineas_no_canonicas,
            "lineas_base_directa": lineas_base_directa,
            "lineas_base_directa_sugeridas": lineas_base_directa_sugeridas,
            "alertas_capacidad": alertas_capacidad,
            "master_incompletos": master_incompletos,
            "health_label": health_label,
            "health_tone": health_tone,
            "health_detail": health_detail,
            "quality_cards": quality_cards,
            "chain_cards": chain_cards,
            "chain_detail_rows": filtered_chain_rows,
            "master_blocker_class_cards": master_blocker_class_cards[:6],
            "master_blocker_missing_cards": master_blocker_missing_cards[:6],
            "master_blocker_detail_rows": filtered_master_rows,
            "master_focus": master_focus,
            "master_focus_rows": master_focus_rows,
            "critical_master_rows": master_focus_rows[:3] if master_demand_critical else [],
            "downstream_handoff_rows": downstream_handoff_rows,
            "demand_signal": demand_signal,
            "commercial_gate": commercial_gate,
            "upstream_dependency_rows": upstream_dependency_rows,
            "selected_focus_kind": selected_focus_kind,
            "selected_focus_key": selected_focus_key,
            "focus_summary": focus_summary,
        }

    critical_path_rows = _recipes_critical_path_rows(
        workflow_rows,
        owner="Producción / Costeo",
        fallback_url=reverse("recetas:mrp_form"),
    )
    return render(
        request,
        "recetas/mrp.html",
        {
            "recetas": recetas,
            "resultado": resultado,
            "erp_command_center": erp_command_center,
            "workflow_rows": workflow_rows,
            "critical_path_rows": critical_path_rows,
        },
    )
