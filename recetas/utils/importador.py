import hashlib
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from django.db import transaction
from django.utils import timezone

from maestros.models import Insumo, CostoInsumo, Proveedor, UnidadMedida, seed_unidades_basicas
from .matching import match_insumo, clasificar_match
from .normalizacion import normalizar_nombre
from .subsection_costing import find_parent_cost_for_stage
from .costeo_versionado import asegurar_version_costeo
from recetas.models import Receta, LineaReceta, RecetaPresentacion

log = logging.getLogger(__name__)

@dataclass
class ImportResultado:
    catalogo_importado: int = 0
    insumos_creados: int = 0
    costos_creados: int = 0
    recetas_creadas: int = 0
    recetas_actualizadas: int = 0
    lineas_creadas: int = 0
    errores: List[Dict[str, Any]] = field(default_factory=list)
    matches_pendientes: List[Dict[str, Any]] = field(default_factory=list)

def _sha256(texto: str) -> str:
    return hashlib.sha256(texto.encode("utf-8")).hexdigest()

def _to_float(x) -> Optional[float]:
    try:
        if x is None:
            return None
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        if s == "":
            return None
        return float(s)
    except Exception:
        return None

def _to_date(x) -> Optional[date]:
    if x is None:
        return None
    if isinstance(x, date) and not isinstance(x, datetime):
        return x
    if isinstance(x, datetime):
        return x.date()
    try:
        return datetime.fromisoformat(str(x)).date()
    except Exception:
        return None

def _normalize_header(val: Any) -> str:
    return normalizar_nombre(val)

def _find_header_row(values: List[List[Any]], required_tokens: List[str], max_scan: int = 30) -> Optional[int]:
    required_tokens = [normalizar_nombre(t) for t in required_tokens]
    for i, row in enumerate(values[:max_scan]):
        row_tokens = " | ".join([_normalize_header(v) for v in row if v is not None])
        if all(t in row_tokens for t in required_tokens):
            return i
    return None

def _map_columns(header_row: List[Any]) -> Dict[str, int]:
    # maps known fields to column index
    mapping = {}
    for idx, cell in enumerate(header_row):
        h = _normalize_header(cell)
        if not h:
            continue
        if h in ("proveedor",):
            mapping["proveedor"] = idx
        elif "producto" in h:
            mapping["producto"] = idx
        elif "descripcion" in h:
            mapping["descripcion"] = idx
        elif "cantidad" == h:
            mapping["cantidad"] = idx
        elif "unidad" == h:
            mapping["unidad"] = idx
        elif h in ("costo", "costounitario", "costo unitario"):
            mapping["costo"] = idx
        elif "$/unidad" in str(cell).lower() or "unidad" == h and "costo" not in mapping:
            # no-op; handled by unidad
            pass
        elif "fecha" in h:
            mapping["fecha"] = idx
        elif "codigo" in h and "barras" in h:
            mapping["codigo_barras"] = idx
        elif "precio" in h:
            mapping["precio"] = idx
    return mapping

def _get_unit(code: str) -> Optional[UnidadMedida]:
    if not code:
        return None
    code_norm = normalizar_nombre(code)
    # unify
    repl = {
        "pz": "pza",
        "pieza": "pza",
        "pzas": "pza",
        "kg": "kg",
        "g": "g",
        "gr": "g",
        "gramo": "g",
        "lt": "lt",
        "l": "lt",
        "litro": "lt",
        "ml": "ml",
    }
    code2 = repl.get(code_norm, code_norm)
    return UnidadMedida.objects.filter(codigo=code2).first()

def _latest_cost_unitario(insumo: Insumo) -> Optional[float]:
    ci = CostoInsumo.objects.filter(insumo=insumo).order_by("-fecha", "-id").first()
    return float(ci.costo_unitario) if ci else None


def _is_presentation_header(text: Any) -> bool:
    n = normalizar_nombre(text)
    return n in {
        "mini",
        "chico",
        "mediano",
        "grande",
        "individual",
        "rebanada",
        "bollo",
        "bollos",
        "bollito",
        "rosca",
        "media plancha",
        "1/2 plancha",
        "1 2 plancha",
    }


def _presentation_display_name(text: Any) -> str:
    n = normalizar_nombre(text)
    if n in {"1/2 plancha", "1 2 plancha", "media plancha"}:
        return "Media Plancha"
    mapping = {
        "mini": "Mini",
        "chico": "Chico",
        "mediano": "Mediano",
        "grande": "Grande",
        "individual": "Individual",
        "rebanada": "Rebanada",
        "bollo": "Bollo",
        "bollos": "Bollos",
        "bollito": "Bollito",
        "rosca": "Rosca",
    }
    return mapping.get(n, str(text).strip())


def _to_positive_decimal(value: Any) -> Optional[Decimal]:
    f = _to_float(value)
    if f is None or f <= 0:
        return None
    return Decimal(str(f)).quantize(Decimal("0.000001"))


_CELL_REF_RE = re.compile(
    r"(?:'(?P<sheet1>[^']+)'|(?P<sheet2>[A-Za-z0-9 _\-\[\]]+))!\$?(?P<col>[A-Z]{1,3})\$?(?P<row>\d+)"
)


def _build_insumos2_pan_map(ws_insumos2) -> Dict[int, str]:
    row_to_pan: Dict[int, str] = {}
    current_pan = ""
    max_row = min(ws_insumos2.max_row, 500)
    for r in range(1, max_row + 1):
        v9 = ws_insumos2.cell(row=r, column=9).value
        v19 = ws_insumos2.cell(row=r, column=19).value
        candidate = v9 if isinstance(v9, str) and v9.strip() else v19
        if isinstance(candidate, str) and candidate.strip():
            n = normalizar_nombre(candidate)
            if n.startswith("pan "):
                current_pan = candidate.strip()
        if current_pan:
            row_to_pan[r] = current_pan
    return row_to_pan


def _resolve_pan_from_formula(formula: Any, pan_map: Dict[int, str]) -> str:
    if not isinstance(formula, str):
        return ""
    m = _CELL_REF_RE.search(formula)
    if not m:
        return ""
    sheet = (m.group("sheet1") or m.group("sheet2") or "").strip()
    if normalizar_nombre(sheet) != "insumos 2":
        return ""
    try:
        row = int(m.group("row"))
    except Exception:
        return ""
    return pan_map.get(row, "")


def _should_autocreate_component(recipe_type: str, tipo_linea: str, ingrediente: str, costo: Optional[float]) -> bool:
    if recipe_type != Receta.TIPO_PRODUCTO_FINAL:
        return False
    if tipo_linea != LineaReceta.TIPO_NORMAL:
        return False
    if costo is None or costo <= 0:
        return False
    ingrediente_norm = normalizar_nombre(ingrediente)
    if not ingrediente_norm:
        return False
    # Evitar crear conceptos operativos que no son insumo real.
    if ingrediente_norm in {"armado", "presentacion", "presentación"}:
        return False
    return ("-" in ingrediente) or (len(ingrediente_norm.split()) >= 2)


def _get_or_create_component_insumo(nombre: str, unidad: Optional[UnidadMedida]) -> Insumo:
    nombre_norm = normalizar_nombre(nombre)
    insumo = Insumo.objects.filter(nombre_normalizado=nombre_norm).order_by("id").first()
    if insumo:
        if insumo.unidad_base is None and unidad is not None:
            insumo.unidad_base = unidad
            insumo.save(update_fields=["unidad_base"])
        return insumo
    return Insumo.objects.create(nombre=nombre[:250], unidad_base=unidad)


def _hydrate_subsection_costs(lineas: List[Dict[str, Any]]) -> None:
    main_costs: List[Tuple[str, float]] = []
    for l in lineas:
        if l.get("tipo_linea") != LineaReceta.TIPO_NORMAL:
            continue
        costo = _to_float(l.get("costo"))
        if costo is None or costo <= 0:
            continue
        main_costs.append((str(l.get("ingrediente") or ""), costo))

    if not main_costs:
        return

    by_stage: Dict[str, List[Dict[str, Any]]] = {}
    for l in lineas:
        if l.get("tipo_linea") != LineaReceta.TIPO_SUBSECCION:
            continue
        if _to_float(l.get("costo")) not in (None, 0):
            continue
        qty = _to_float(l.get("cantidad"))
        if qty is None or qty <= 0:
            continue
        stage = str(l.get("etapa") or "").strip()
        if not stage:
            continue
        by_stage.setdefault(stage, []).append(l)

    for stage, stage_lines in by_stage.items():
        stage_total_cost = None
        for l in stage_lines:
            v = _to_float(l.get("_stage_total_cost"))
            if v is not None and v > 0:
                stage_total_cost = v
                break
        parent_cost = stage_total_cost if stage_total_cost else find_parent_cost_for_stage(stage, main_costs)
        if parent_cost is None or parent_cost <= 0:
            continue
        total_qty = sum((_to_float(l.get("cantidad")) or 0.0) for l in stage_lines)
        if total_qty <= 0:
            continue
        for l in stage_lines:
            qty = _to_float(l.get("cantidad")) or 0.0
            l["costo"] = parent_cost * (qty / total_qty)


def _ensure_unique_receta_hash(desired_hash: str, receta_id: Optional[int], seed: str) -> str:
    candidate = desired_hash
    i = 0
    while Receta.objects.filter(hash_contenido=candidate).exclude(pk=receta_id).exists():
        i += 1
        candidate = _sha256(f"{desired_hash}|{seed}|{i}")
    return candidate


def _remove_duplicate_recetas(nombre_norm: str, keep_id: int) -> None:
    Receta.objects.filter(nombre_normalizado=nombre_norm).exclude(pk=keep_id).delete()


class ImportadorCosteo:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(filepath, data_only=True)
        self.wb_formula = openpyxl.load_workbook(filepath, data_only=False)
        self.pan_map: Dict[int, str] = {}
        if "Insumos 2" in self.wb_formula.sheetnames:
            self.pan_map = _build_insumos2_pan_map(self.wb_formula["Insumos 2"])
        self.resultado = ImportResultado()

    @transaction.atomic
    def procesar_completo(self) -> ImportResultado:
        seed_unidades_basicas()
        self.importar_catalogo_costos()
        self.importar_recetas()
        return self.resultado

    def importar_catalogo_costos(self):
        if "Costo Materia Prima" not in self.wb.sheetnames:
            self.resultado.errores.append({"sheet": "Costo Materia Prima", "error": "No existe la hoja"})
            return

        ws = self.wb["Costo Materia Prima"]
        values = list(ws.values)
        header_i = _find_header_row(values, ["Proveedor", "Producto", "Costo"])
        if header_i is None:
            header_i = 1  # fallback
        header = list(values[header_i])
        colmap = _map_columns(header)

        # data rows
        for row in values[header_i + 1 :]:
            try:
                producto = row[colmap.get("producto", 3)] if len(row) > colmap.get("producto", 3) else None
                desc = row[colmap.get("descripcion", 4)] if len(row) > colmap.get("descripcion", 4) else None
                nombre = (producto or desc)
                if not nombre:
                    continue

                proveedor_nombre = None
                if "proveedor" in colmap and len(row) > colmap["proveedor"]:
                    proveedor_nombre = row[colmap["proveedor"]]
                proveedor = None
                if proveedor_nombre:
                    proveedor, _ = Proveedor.objects.get_or_create(nombre=str(proveedor_nombre).strip()[:200])

                unidad_code = None
                if "unidad" in colmap and len(row) > colmap["unidad"]:
                    unidad_code = row[colmap["unidad"]]
                # en este excel, la unidad real puede estar duplicada; preferimos Unidad
                unidad = _get_unit(str(unidad_code)) if unidad_code else None

                costo = None
                if "costo" in colmap and len(row) > colmap["costo"]:
                    costo = _to_float(row[colmap["costo"]])

                cantidad = None
                if "cantidad" in colmap and len(row) > colmap["cantidad"]:
                    cantidad = _to_float(row[colmap["cantidad"]])

                precio = None
                if "precio" in colmap and len(row) > colmap["precio"]:
                    precio = _to_float(row[colmap["precio"]])

                fecha = None
                if "fecha" in colmap and len(row) > colmap["fecha"]:
                    fecha = _to_date(row[colmap["fecha"]])

                # Si costo no existe pero precio y cantidad sí, intentar calcular
                costo_unitario = costo
                if costo_unitario is None and precio is not None and cantidad and cantidad > 0:
                    costo_unitario = precio / cantidad

                if costo_unitario is None:
                    continue

                insumo, created = Insumo.objects.get_or_create(
                    nombre_normalizado=normalizar_nombre(nombre),
                    defaults={
                        "nombre": str(nombre).strip()[:250],
                        "unidad_base": unidad,
                        "proveedor_principal": proveedor,
                    },
                )
                if created:
                    self.resultado.insumos_creados += 1
                else:
                    # actualizar unidad/proveedor si no tiene
                    changed = False
                    if insumo.unidad_base is None and unidad is not None:
                        insumo.unidad_base = unidad
                        changed = True
                    if insumo.proveedor_principal is None and proveedor is not None:
                        insumo.proveedor_principal = proveedor
                        changed = True
                    if changed:
                        insumo.save()

                raw = {
                    "producto": str(producto) if producto is not None else "",
                    "descripcion": str(desc) if desc is not None else "",
                    "precio": precio,
                    "cantidad": cantidad,
                    "unidad": str(unidad_code) if unidad_code is not None else "",
                    "fecha": str(fecha) if fecha else "",
                    "proveedor": str(proveedor_nombre) if proveedor_nombre else "",
                }
                source_hash = _sha256(json_dumps_sorted({"n": insumo.nombre_normalizado, "c": costo_unitario, "f": str(fecha), "p": raw.get("proveedor","")}))
                if not CostoInsumo.objects.filter(source_hash=source_hash).exists():
                    CostoInsumo.objects.create(
                        insumo=insumo,
                        proveedor=proveedor,
                        fecha=fecha or timezone.now().date(),
                        costo_unitario=costo_unitario,
                        source_hash=source_hash,
                        raw=raw,
                    )
                    self.resultado.costos_creados += 1
            except Exception as e:
                self.resultado.errores.append({"sheet": "Costo Materia Prima", "error": str(e)})

        self.resultado.catalogo_importado = 1

    def importar_recetas(self):
        # detecta hojas con recetas tradicionales y/o matrices por presentación.
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            has_ingredientes = self._sheet_has_ingredientes(ws)
            has_matrix = self._sheet_has_presentacion_matrix(ws)
            if not has_ingredientes and not has_matrix:
                continue
            try:
                if has_ingredientes:
                    self._importar_recetas_de_hoja(ws, sheet_name)
                if has_matrix:
                    self._importar_productos_finales_de_hoja(ws, sheet_name)
            except Exception as e:
                self.resultado.errores.append({"sheet": sheet_name, "error": str(e)})

    def _sheet_has_ingredientes(self, ws) -> bool:
        for r in range(1, 80):
            for c in range(1, 8):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and "ingrediente" in v.lower():
                    return True
        return False

    def _sheet_has_presentacion_matrix(self, ws) -> bool:
        if "pastel" not in normalizar_nombre(ws.title):
            return False
        max_row = min(ws.max_row, 120)
        max_col = min(ws.max_column, 24)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                if normalizar_nombre(ws.cell(row=r, column=c).value) != "elemento":
                    continue
                for cc in range(c + 1, min(c + 8, max_col + 1)):
                    if _is_presentation_header(ws.cell(row=r, column=cc).value):
                        return True
        return False

    def _infer_rendimiento_unidad(self, lineas: List[Dict[str, Any]], fg_hint: str) -> Optional[UnidadMedida]:
        if fg_hint:
            u = _get_unit(fg_hint)
            if u:
                return u

        counts: Dict[str, int] = {"kg": 0, "lt": 0}
        for l in lineas:
            u = normalizar_nombre(l.get("unidad", ""))
            if u in {"kg", "g", "gr"}:
                counts["kg"] += 1
            elif u in {"lt", "l", "litro", "ml"}:
                counts["lt"] += 1
        if counts["lt"] > counts["kg"] and counts["lt"] > 0:
            return _get_unit("lt")
        if counts["kg"] > 0:
            return _get_unit("kg")
        return None

    def _extract_presentaciones_block(self, ws, start_row: int, end_row: int) -> List[Dict[str, Any]]:
        scan_end = min(ws.max_row, end_row)
        found: Dict[str, Dict[str, Any]] = {}

        for r in range(start_row, scan_end + 1):
            headers: Dict[int, str] = {}
            for c in range(9, 18):
                hv = ws.cell(row=r, column=c).value
                if _is_presentation_header(hv):
                    headers[c] = _presentation_display_name(hv)
            if not headers:
                continue

            # Patrón simple: fila de encabezados y siguiente fila con pesos.
            simple_weight_row = r + 1 if r + 1 <= scan_end else None
            for c, name in headers.items():
                weight = _to_float(ws.cell(row=simple_weight_row, column=c).value) if simple_weight_row else None
                if weight is None or weight <= 0:
                    continue
                key = normalizar_nombre(name)
                found[key] = {
                    "nombre": name[:80],
                    "peso_por_unidad_kg": weight,
                    "unidades_por_batch": None,
                    "unidades_por_pastel": None,
                }

        # Patrón con etiquetas en columna I: Elemento / Peso-Molde / Cantidades.
        for r in range(start_row, scan_end + 1):
            if normalizar_nombre(ws.cell(row=r, column=9).value) != "elemento":
                continue

            headers: Dict[int, str] = {}
            for c in range(10, 18):
                hv = ws.cell(row=r, column=c).value
                if _is_presentation_header(hv):
                    headers[c] = _presentation_display_name(hv)
            if not headers:
                continue

            weight_row = None
            batch_row = None
            pastel_row = None
            for rr in range(r + 1, min(scan_end, r + 8) + 1):
                label = normalizar_nombre(ws.cell(row=rr, column=9).value)
                if label in {"peso/molde", "peso molde", "peso por molde"}:
                    weight_row = rr
                elif label in {"cantidad piezas en pan", "cantidad piezas pan", "cantidad piezas"}:
                    batch_row = rr
                elif label in {"cantidad en pasteles", "cantidad en pastel"}:
                    pastel_row = rr

            for c, name in headers.items():
                weight = _to_float(ws.cell(row=weight_row, column=c).value) if weight_row else None
                if weight is None or weight <= 0:
                    continue
                key = normalizar_nombre(name)
                item = found.get(
                    key,
                    {
                        "nombre": name[:80],
                        "peso_por_unidad_kg": weight,
                        "unidades_por_batch": None,
                        "unidades_por_pastel": None,
                    },
                )
                item["peso_por_unidad_kg"] = weight
                if batch_row:
                    item["unidades_por_batch"] = _to_positive_decimal(ws.cell(row=batch_row, column=c).value)
                if pastel_row:
                    item["unidades_por_pastel"] = _to_positive_decimal(ws.cell(row=pastel_row, column=c).value)
                found[key] = item

        return list(found.values())

    def _extract_preparacion_metadata(
        self,
        ws,
        start_row: int,
        end_row: int,
        lineas: List[Dict[str, Any]],
    ) -> Tuple[Optional[float], Optional[UnidadMedida], List[Dict[str, Any]]]:
        rendimiento_cantidad = None
        costo_unitario_fg = None
        unit_hint = ""
        fallback_qty = None
        fallback_unit_hint = ""
        scan_end = min(ws.max_row, end_row + 30)

        for r in range(start_row, scan_end + 1):
            # Evitar capturar metadatos de la siguiente receta.
            if r > end_row:
                maybe_recipe_title = ws.cell(row=r, column=1).value
                maybe_header = ws.cell(row=r + 1, column=1).value if (r + 1) <= ws.max_row else None
                if (
                    isinstance(maybe_recipe_title, str)
                    and maybe_recipe_title.strip()
                    and isinstance(maybe_header, str)
                    and "ingrediente" in maybe_header.lower()
                ):
                    break

            c1 = ws.cell(row=r, column=1).value
            c2 = ws.cell(row=r, column=2).value
            c3 = ws.cell(row=r, column=3).value
            c4 = ws.cell(row=r, column=4).value
            c5 = ws.cell(row=r, column=5).value
            f_raw = ws.cell(row=r, column=6).value
            g_val = _to_float(ws.cell(row=r, column=7).value)
            f_norm = normalizar_nombre(f_raw)
            c1_norm = normalizar_nombre(c1)
            if g_val is not None and g_val > 0:
                is_cost_row = ("costo" in f_norm) or ("$/" in f_norm) or ("$/ " in f_norm)
                looks_like_total = any(token in f_norm for token in {"batida", "total", "peso", "rendimiento"})

                if (
                    f_norm in {"batida", "total batida", "total peso", "peso total", "total"}
                    or (looks_like_total and not is_cost_row)
                ):
                    rendimiento_cantidad = g_val

                if is_cost_row:
                    costo_unitario_fg = g_val

            # Fallback flexible fuera de F/G:
            # detecta filas tipo "Rendimiento", "Rendimiento Individuales", "Numero de galletas".
            if any(t in c1_norm for t in {"rendimiento", "numero de galletas", "numero galletas"}):
                qty_candidate = None
                unit_candidate = ""

                c2f = _to_float(c2)
                c3f = _to_float(c3)
                c4f = _to_float(c4)
                c5f = _to_float(c5)

                unit_from_c2 = _get_unit(str(c2)) if isinstance(c2, str) else None
                unit_from_c3 = _get_unit(str(c3)) if isinstance(c3, str) else None
                unit_from_c4 = _get_unit(str(c4)) if isinstance(c4, str) else None

                if c2f is not None and c2f > 0 and unit_from_c3:
                    qty_candidate = c2f
                    unit_candidate = unit_from_c3.codigo
                elif c3f is not None and c3f > 0:
                    qty_candidate = c3f
                    if unit_from_c2:
                        unit_candidate = unit_from_c2.codigo
                elif c4f is not None and c4f > 0:
                    qty_candidate = c4f
                    if unit_from_c3:
                        unit_candidate = unit_from_c3.codigo
                    elif unit_from_c2:
                        unit_candidate = unit_from_c2.codigo
                elif c5f is not None and c5f > 0:
                    qty_candidate = c5f
                    if unit_from_c4:
                        unit_candidate = unit_from_c4.codigo

                if qty_candidate is not None and qty_candidate > 0:
                    fallback_qty = qty_candidate
                    if not unit_candidate:
                        # Si no hay unidad explícita, para rendimiento por piezas default a pza.
                        if any(t in c1_norm for t in {"galleta", "individual", "empanada", "pieza", "pza", "numero"}):
                            unit_candidate = "pza"
                        elif "rendimiento" in c1_norm:
                            unit_candidate = "pza"
                    if unit_candidate:
                        fallback_unit_hint = unit_candidate

            if f_raw is not None:
                f_txt = str(f_raw).lower()
                if ("kg" in f_txt) and ("costo" in f_txt or "$/" in f_txt):
                    unit_hint = "kg"
                elif ("lt" in f_txt or "/l" in f_txt or "litro" in f_txt) and ("costo" in f_txt or "$/" in f_txt):
                    unit_hint = "lt"

        if (rendimiento_cantidad is None or rendimiento_cantidad <= 0) and costo_unitario_fg and costo_unitario_fg > 0:
            total_costo_lineas = sum((_to_float(l.get("costo")) or 0.0) for l in lineas)
            if total_costo_lineas > 0:
                rendimiento_cantidad = total_costo_lineas / costo_unitario_fg
        if (rendimiento_cantidad is None or rendimiento_cantidad <= 0) and fallback_qty and fallback_qty > 0:
            rendimiento_cantidad = fallback_qty
        if not unit_hint and fallback_unit_hint:
            unit_hint = fallback_unit_hint

        rendimiento_unidad = self._infer_rendimiento_unidad(lineas, unit_hint)
        presentaciones = self._extract_presentaciones_block(ws, start_row, end_row)
        return rendimiento_cantidad, rendimiento_unidad, presentaciones

    def _importar_recetas_de_hoja(self, ws, sheet_name: str):
        max_row = min(ws.max_row, 400)
        r = 1
        while r <= max_row - 2:
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r+1, column=1).value
            if isinstance(a, str) and a.strip() and isinstance(b, str) and "ingrediente" in b.lower():
                receta_nombre = a.strip()
                header_row = r + 1
                # detect columns based on header row
                header_vals = [ws.cell(row=header_row, column=c).value for c in range(1, 10)]
                col_ing = 1
                col_qty = 2
                col_unit = 3
                col_cost = 4
                for idx, hv in enumerate(header_vals, start=1):
                    hvn = normalizar_nombre(hv)
                    if hvn in ("ingredientes", "insumo", "ingrediente"):
                        col_ing = idx
                    elif hvn.startswith("cantidad"):
                        col_qty = idx
                    elif hvn.startswith("unidad"):
                        col_unit = idx
                    elif "costo" in hvn or hvn == "$":
                        col_cost = idx

                lineas = []
                rr = header_row + 1
                pos = 1
                while rr <= max_row:
                    ing = ws.cell(row=rr, column=col_ing).value
                    qty = ws.cell(row=rr, column=col_qty).value if col_qty else None
                    unit = ws.cell(row=rr, column=col_unit).value if col_unit else None
                    cost = ws.cell(row=rr, column=col_cost).value if col_cost else None

                    # stop conditions
                    if isinstance(ing, str) and normalizar_nombre(ing) in ("total", "costo total"):
                        break
                    if isinstance(qty, str) and "costo total" in qty.lower():
                        break
                    if ing is None and (qty is None or (isinstance(qty, str) and qty.strip()=="")):
                        # blank row => end block
                        break
                    if isinstance(ing, str) and "costo total" in ing.lower():
                        break
                    if isinstance(qty, str) and normalizar_nombre(qty) == "total":
                        break

                    if ing is not None and str(ing).strip():
                        lineas.append({
                            "pos": pos,
                            "tipo_linea": LineaReceta.TIPO_NORMAL,
                            "etapa": "",
                            "ingrediente": str(ing).strip(),
                            "cantidad": _to_float(qty),
                            "unidad": str(unit).strip() if unit is not None else "",
                            "costo": _to_float(cost),
                        })
                        pos += 1
                    rr += 1

                if lineas:
                    rendimiento_cantidad, rendimiento_unidad, presentaciones = self._extract_preparacion_metadata(
                        ws,
                        r,
                        rr,
                        lineas,
                    )
                    self._upsert_receta(
                        receta_nombre,
                        sheet_name,
                        lineas,
                        recipe_type=Receta.TIPO_PREPARACION,
                        rendimiento_cantidad=rendimiento_cantidad,
                        rendimiento_unidad=rendimiento_unidad,
                        presentaciones=presentaciones,
                    )
                    r = rr + 1
                    continue

            r += 1

    def _importar_productos_finales_de_hoja(self, ws, sheet_name: str):
        if "pastel" not in normalizar_nombre(sheet_name):
            return
        max_row = min(ws.max_row, 250)
        max_col = min(ws.max_column, 30)
        ws_formula = self.wb_formula[sheet_name] if sheet_name in self.wb_formula.sheetnames else ws
        values = [
            [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            for r in range(1, max_row + 1)
        ]
        formula_values = [
            [ws_formula.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            for r in range(1, max_row + 1)
        ]

        def v(r: int, c: int):
            return values[r - 1][c - 1]

        def vf(r: int, c: int):
            return formula_values[r - 1][c - 1]

        recetas_por_presentacion: Dict[str, List[Dict[str, Any]]] = {}
        size_cols: Dict[str, Tuple[int, str]] = {}

        # Bloque principal de costos por componente (columna A + tamaños en la fila de encabezado).
        for r in range(1, max_row + 1):
            head = normalizar_nombre(v(r, 1))
            if head not in {"insumo", "ingrediente", "insumos"}:
                continue
            headers_started = False
            for c in range(2, max_col + 1):
                hv = v(r, c)
                if hv is None or str(hv).strip() == "":
                    if headers_started:
                        break
                    continue
                if _is_presentation_header(hv):
                    size_cols[normalizar_nombre(hv)] = (c, str(hv).strip())
                    headers_started = True
                elif headers_started:
                    break

            if size_cols:
                rr = r + 1
                while rr <= max_row:
                    componente = v(rr, 1)
                    if componente is None or str(componente).strip() == "":
                        rr += 1
                        continue
                    componente_txt = str(componente).strip()
                    componente_norm = normalizar_nombre(componente_txt)
                    if componente_norm.startswith("subtotal") or componente_norm.startswith("costo"):
                        break

                    for _, (col, presentacion) in size_cols.items():
                        costo = _to_float(v(rr, col))
                        if costo is None or costo <= 0:
                            continue
                        ingrediente_txt = componente_txt
                        if componente_norm == "pan":
                            pan_name = _resolve_pan_from_formula(vf(rr, col), self.pan_map)
                            if pan_name:
                                ingrediente_txt = f"{pan_name} - {presentacion}"
                        receta_nombre = f"{sheet_name} - {presentacion}"[:250]
                        lineas = recetas_por_presentacion.setdefault(receta_nombre, [])
                        lineas.append(
                            {
                                "pos": len(lineas) + 1,
                                "tipo_linea": LineaReceta.TIPO_NORMAL,
                                "etapa": "",
                                "ingrediente": ingrediente_txt[:250],
                                "cantidad": None,
                                "unidad": "",
                                "costo": costo,
                            }
                        )
                    rr += 1
            break

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                if normalizar_nombre(v(r, c)) != "elemento":
                    continue

                headers: List[Tuple[int, str]] = []
                cc = c + 1
                while cc <= max_col:
                    hv = v(r, cc)
                    if hv is None or str(hv).strip() == "":
                        if headers:
                            break
                        cc += 1
                        continue
                    if isinstance(hv, (int, float)):
                        break
                    if _is_presentation_header(hv):
                        headers.append((cc, str(hv).strip()))
                    cc += 1

                if not headers:
                    continue

                section = ""
                if r > 1:
                    prev = v(r - 1, c)
                    if isinstance(prev, str) and prev.strip():
                        section = prev.strip()[:120]

                section_items: List[Tuple[str, Dict[str, float]]] = []
                rr = r + 1
                while rr <= max_row:
                    elemento = v(rr, c)
                    if elemento is None or str(elemento).strip() == "":
                        break
                    elemento_txt = str(elemento).strip()
                    elemento_norm = normalizar_nombre(elemento_txt)
                    if elemento_norm == "elemento" or elemento_norm.startswith("total"):
                        break

                    qty_by_presentation: Dict[str, float] = {}
                    for col, presentacion in headers:
                        cantidad = _to_float(v(rr, col))
                        if cantidad is None or cantidad <= 0:
                            continue
                        qty_by_presentation[presentacion] = cantidad
                    if qty_by_presentation:
                        section_items.append((elemento_txt[:250], qty_by_presentation))
                    rr += 1

                section_total_by_presentation: Dict[str, float] = {}
                scan = rr
                while scan <= max_row:
                    label = v(scan, c)
                    if label is None or str(label).strip() == "":
                        if scan - rr > 6:
                            break
                        scan += 1
                        continue

                    label_norm = normalizar_nombre(label)
                    if label_norm == "total por decorar":
                        for col, presentacion in headers:
                            total_cost = _to_float(v(scan, col))
                            if total_cost is not None and total_cost > 0:
                                section_total_by_presentation[presentacion] = total_cost
                        break
                    if label_norm == "elemento":
                        break
                    if scan + 1 <= max_row and normalizar_nombre(v(scan + 1, c)) == "elemento":
                        break
                    scan += 1

                totals_qty_by_presentation: Dict[str, float] = {}
                for _, qty_map in section_items:
                    for presentacion, qty in qty_map.items():
                        totals_qty_by_presentation[presentacion] = totals_qty_by_presentation.get(presentacion, 0.0) + qty

                for elemento_txt, qty_map in section_items:
                    for presentacion, cantidad in qty_map.items():
                        receta_nombre = f"{sheet_name} - {presentacion}"[:250]
                        lineas = recetas_por_presentacion.setdefault(receta_nombre, [])
                        stage_total = section_total_by_presentation.get(presentacion)
                        total_qty = totals_qty_by_presentation.get(presentacion, 0.0)
                        allocated = None
                        if stage_total is not None and stage_total > 0 and total_qty > 0:
                            allocated = stage_total * (cantidad / total_qty)
                        lineas.append(
                            {
                                "pos": len(lineas) + 1,
                                "tipo_linea": LineaReceta.TIPO_SUBSECCION,
                                "etapa": section,
                                "ingrediente": elemento_txt[:250],
                                "cantidad": cantidad,
                                "unidad": "kg",
                                "costo": allocated,
                                "_stage_total_cost": stage_total,
                            }
                        )

        for receta_nombre, lineas in recetas_por_presentacion.items():
            if not lineas:
                continue
            _hydrate_subsection_costs(lineas)
            self._upsert_receta(
                receta_nombre,
                sheet_name,
                lineas,
                recipe_type=Receta.TIPO_PRODUCTO_FINAL,
            )

    def _upsert_receta(
        self,
        nombre: str,
        sheet_name: str,
        lineas: List[Dict[str, Any]],
        recipe_type: str = Receta.TIPO_PREPARACION,
        rendimiento_cantidad: Optional[float] = None,
        rendimiento_unidad: Optional[UnidadMedida] = None,
        presentaciones: Optional[List[Dict[str, Any]]] = None,
    ):
        # Hash idempotente por contenido
        nombre_norm = normalizar_nombre(nombre)
        h_payload = {
            "nombre": nombre_norm,
            "sheet": sheet_name,
            "lineas": [
                (
                    normalizar_nombre(l["ingrediente"]),
                    l.get("cantidad"),
                    normalizar_nombre(l.get("unidad", "")),
                    l.get("tipo_linea", LineaReceta.TIPO_NORMAL),
                    normalizar_nombre(l.get("etapa", "")),
                )
                for l in lineas
            ],
        }
        desired_hash = _sha256(json_dumps_sorted(h_payload))

        # Consolidar por nombre normalizado para evitar duplicados entre corridas.
        receta = Receta.objects.filter(nombre_normalizado=nombre_norm).order_by("-id").first()
        if not receta:
            receta = Receta.objects.filter(hash_contenido=desired_hash).first()
        hash_contenido = _ensure_unique_receta_hash(
            desired_hash,
            receta.id if receta else None,
            nombre_norm,
        )
        if receta:
            self.resultado.recetas_actualizadas += 1
            # asegurar nombre/sheet/hash y recrear lineas
            changed = False
            if (
                receta.nombre != nombre
                or receta.sheet_name != sheet_name
                or receta.hash_contenido != hash_contenido
                or receta.tipo != recipe_type
            ):
                receta.nombre = nombre
                receta.sheet_name = sheet_name
                receta.hash_contenido = hash_contenido
                receta.tipo = recipe_type
                changed = True

            if rendimiento_cantidad is not None and rendimiento_cantidad > 0:
                qty = Decimal(str(rendimiento_cantidad))
                if receta.rendimiento_cantidad != qty:
                    receta.rendimiento_cantidad = qty
                    changed = True
            if rendimiento_unidad is not None and receta.rendimiento_unidad_id != rendimiento_unidad.id:
                receta.rendimiento_unidad = rendimiento_unidad
                changed = True

            if presentaciones is not None and recipe_type == Receta.TIPO_PREPARACION:
                usa_presentaciones = bool(presentaciones)
                if receta.usa_presentaciones != usa_presentaciones:
                    receta.usa_presentaciones = usa_presentaciones
                    changed = True

            if changed:
                receta.save()
            _remove_duplicate_recetas(nombre_norm, receta.id)
            receta.lineas.all().delete()
        else:
            receta = Receta.objects.create(
                nombre=nombre,
                sheet_name=sheet_name,
                hash_contenido=hash_contenido,
                tipo=recipe_type,
                rendimiento_cantidad=Decimal(str(rendimiento_cantidad)) if rendimiento_cantidad else None,
                rendimiento_unidad=rendimiento_unidad,
                usa_presentaciones=bool(presentaciones) if (presentaciones is not None and recipe_type == Receta.TIPO_PREPARACION) else False,
            )
            self.resultado.recetas_creadas += 1
            _remove_duplicate_recetas(nombre_norm, receta.id)

        for l in lineas:
            insumo, score, method = match_insumo(l["ingrediente"])
            status = clasificar_match(score)
            unidad = _get_unit(l.get("unidad",""))
            tipo_linea = l.get("tipo_linea", LineaReceta.TIPO_NORMAL)
            etapa = str(l.get("etapa") or "")[:120]
            if insumo is None and _should_autocreate_component(
                recipe_type=recipe_type,
                tipo_linea=tipo_linea,
                ingrediente=l["ingrediente"],
                costo=l.get("costo"),
            ):
                insumo = _get_or_create_component_insumo(l["ingrediente"], unidad)
                score = 100.0
                method = "AUTO_COMPONENTE"
                status = LineaReceta.STATUS_AUTO
            costo_unitario = _latest_cost_unitario(insumo) if insumo else None
            if tipo_linea == LineaReceta.TIPO_SUBSECCION and insumo is None:
                status = LineaReceta.STATUS_AUTO
                method = LineaReceta.MATCH_SUBSECTION
                score = 100.0
            lr = LineaReceta.objects.create(
                receta=receta,
                posicion=l["pos"],
                tipo_linea=tipo_linea,
                etapa=etapa,
                insumo=insumo if status != "REJECTED" else None,
                insumo_texto=l["ingrediente"],
                cantidad=l.get("cantidad"),
                unidad_texto=l.get("unidad",""),
                unidad=unidad,
                costo_linea_excel=l.get("costo"),
                costo_unitario_snapshot=costo_unitario,
                match_score=score,
                match_method=method,
                match_status=status,
            )
            self.resultado.lineas_creadas += 1
            if status == "NEEDS_REVIEW":
                self.resultado.matches_pendientes.append({
                    "receta": receta.nombre,
                    "ingrediente": l["ingrediente"],
                    "score": round(score, 2),
                    "method": method,
                    "status": status,
                })

        if presentaciones is not None and recipe_type == Receta.TIPO_PREPARACION:
            incoming: Dict[str, Dict[str, Any]] = {}
            for p in presentaciones:
                nombre_p = str(p.get("nombre") or "").strip()[:80]
                peso = _to_float(p.get("peso_por_unidad_kg"))
                if not nombre_p or peso is None or peso <= 0:
                    continue
                incoming[normalizar_nombre(nombre_p)] = {
                    "nombre": nombre_p,
                    "peso_por_unidad_kg": Decimal(str(peso)),
                    "unidades_por_batch": p.get("unidades_por_batch"),
                    "unidades_por_pastel": p.get("unidades_por_pastel"),
                }

            existing = {
                normalizar_nombre(x.nombre): x
                for x in receta.presentaciones.all()
            }

            for key, data in incoming.items():
                obj = existing.get(key)
                if obj:
                    changed_fields: List[str] = []
                    if obj.nombre != data["nombre"]:
                        obj.nombre = data["nombre"]
                        changed_fields.append("nombre")
                    if obj.peso_por_unidad_kg != data["peso_por_unidad_kg"]:
                        obj.peso_por_unidad_kg = data["peso_por_unidad_kg"]
                        changed_fields.append("peso_por_unidad_kg")
                    if obj.unidades_por_batch != data["unidades_por_batch"]:
                        obj.unidades_por_batch = data["unidades_por_batch"]
                        changed_fields.append("unidades_por_batch")
                    if obj.unidades_por_pastel != data["unidades_por_pastel"]:
                        obj.unidades_por_pastel = data["unidades_por_pastel"]
                        changed_fields.append("unidades_por_pastel")
                    if not obj.activo:
                        obj.activo = True
                        changed_fields.append("activo")
                    if changed_fields:
                        obj.save(update_fields=changed_fields)
                else:
                    RecetaPresentacion.objects.create(
                        receta=receta,
                        nombre=data["nombre"],
                        peso_por_unidad_kg=data["peso_por_unidad_kg"],
                        unidades_por_batch=data["unidades_por_batch"],
                        unidades_por_pastel=data["unidades_por_pastel"],
                        activo=True,
                    )

            for key, obj in existing.items():
                if key not in incoming:
                    obj.delete()

        try:
            asegurar_version_costeo(receta, fuente="IMPORT_COSTEO")
        except Exception as exc:
            log.warning("No se pudo versionar receta %s (%s)", receta.id, exc)

def json_dumps_sorted(obj: Any) -> str:
    import json
    return json.dumps(obj, ensure_ascii=False, sort_keys=True, default=str)
