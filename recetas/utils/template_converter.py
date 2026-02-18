from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from recetas.utils.normalizacion import normalizar_nombre


TEMPLATE_HEADERS = [
    "receta",
    "subreceta",
    "producto_final",
    "tipo",
    "tipo_linea",
    "etapa",
    "ingrediente",
    "cantidad",
    "unidad",
    "costo_linea",
    "orden",
    "notas",
]


@dataclass
class ConversionResult:
    recetas_detectadas: int = 0
    lineas_detectadas: int = 0
    hojas_escaneadas: int = 0
    hojas_con_recetas: int = 0


def _to_number(value: Any):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    raw = str(value).strip().replace(",", ".")
    if not raw:
        return ""
    try:
        return float(raw)
    except ValueError:
        return ""


def _find_recipe_block(ws) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    r = 1
    max_row = min(ws.max_row, 400)
    while r <= max_row - 2:
        title = ws.cell(row=r, column=1).value
        next_row_values = [ws.cell(row=r + 1, column=c).value for c in range(1, 20)]
        next_row_text = " | ".join(normalizar_nombre(v) for v in next_row_values if v is not None)
        if not (isinstance(title, str) and title.strip()):
            r += 1
            continue
        if "ingrediente" not in next_row_text:
            r += 1
            continue

        receta_nombre = title.strip()[:250]
        subreceta = ws.title[:120]
        header = [ws.cell(row=r + 1, column=c).value for c in range(1, 20)]
        col_ing = 1
        col_qty = 2
        col_unit = 3
        col_cost = 4
        for idx, hv in enumerate(header, start=1):
            h = normalizar_nombre(hv)
            if h in {"ingrediente", "ingredientes", "insumo"}:
                col_ing = idx
            elif "cantidad" in h:
                col_qty = idx
            elif h.startswith("unidad"):
                col_unit = idx
            elif "costo" in h or h in {"$", "costo/$"}:
                col_cost = idx

        rr = r + 2
        orden = 1
        while rr <= max_row:
            ing = ws.cell(row=rr, column=col_ing).value
            qty = ws.cell(row=rr, column=col_qty).value
            unit = ws.cell(row=rr, column=col_unit).value
            cost = ws.cell(row=rr, column=col_cost).value

            ing_norm = normalizar_nombre(ing)
            qty_norm = normalizar_nombre(qty)

            if ing_norm in {"total", "costo total"} or qty_norm in {"total", "costo total"}:
                break
            if ing is None and (qty is None or str(qty).strip() == ""):
                break
            if not isinstance(ing, str) or not ing.strip():
                rr += 1
                continue

            rows.append(
                {
                    "receta": receta_nombre,
                    "subreceta": subreceta,
                    "producto_final": receta_nombre,
                    "tipo": "PREPARACION",
                    "tipo_linea": "NORMAL",
                    "etapa": "",
                    "ingrediente": ing.strip(),
                    "cantidad": _to_number(qty),
                    "unidad": (str(unit).strip() if unit is not None else ""),
                    "costo_linea": _to_number(cost),
                    "orden": orden,
                    "notas": "",
                }
            )
            orden += 1
            rr += 1

        r = rr + 1
    return rows


def _is_presentation_header(text: Any) -> bool:
    n = normalizar_nombre(text)
    return n in {
        "mini",
        "chico",
        "mediano",
        "grande",
        "individual",
        "rebanada",
        "bollos",
        "bollito",
        "media plancha",
        "1/2 plancha",
        "1 2 plancha",
    }


_CELL_REF_RE = re.compile(
    r"(?:'(?P<sheet1>[^']+)'|(?P<sheet2>[A-Za-z0-9 _\-\[\]]+))!\$?(?P<col>[A-Z]{1,3})\$?(?P<row>\d+)"
)


def _build_insumos2_pan_map(ws_insumos2) -> dict[int, str]:
    row_to_pan: dict[int, str] = {}
    current_pan = ""
    max_row = min(ws_insumos2.max_row, 500)
    for r in range(1, max_row + 1):
        v9 = ws_insumos2.cell(row=r, column=9).value
        v19 = ws_insumos2.cell(row=r, column=19).value
        candidate = v9 if isinstance(v9, str) and v9.strip() else v19
        if isinstance(candidate, str) and candidate.strip():
            n = normalizar_nombre(candidate)
            if n.startswith("pan "):
                current_pan = candidate.strip()
        if current_pan:
            row_to_pan[r] = current_pan
    return row_to_pan


def _resolve_pan_from_formula(formula: Any, pan_map: dict[int, str]) -> str:
    if not isinstance(formula, str):
        return ""
    m = _CELL_REF_RE.search(formula)
    if not m:
        return ""
    sheet = (m.group("sheet1") or m.group("sheet2") or "").strip()
    if normalizar_nombre(sheet) != "insumos 2":
        return ""
    try:
        row = int(m.group("row"))
    except Exception:
        return ""
    return pan_map.get(row, "")


def _find_product_final_matrix(ws, ws_formula=None, pan_map: dict[int, str] | None = None) -> list[dict[str, Any]]:
    if "pastel" not in normalizar_nombre(ws.title):
        return []

    rows: list[dict[str, Any]] = []
    orden_por_receta: dict[str, int] = {}
    max_row = min(ws.max_row, 250)
    max_col = min(ws.max_column, 30)
    values = [
        [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]
    ws_formula = ws_formula or ws
    formula_values = [
        [ws_formula.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]

    def v(r: int, c: int):
        return values[r - 1][c - 1]

    def vf(r: int, c: int):
        return formula_values[r - 1][c - 1]

    size_cols: dict[str, tuple[int, str]] = {}

    # Bloque principal: costos por componente (Pan, Betún, Fresa, etc.) por tamaño.
    for r in range(1, max_row + 1):
        head = normalizar_nombre(v(r, 1))
        if head not in {"insumo", "ingrediente", "insumos"}:
            continue
        for c in range(2, max_col + 1):
            hv = v(r, c)
            if not _is_presentation_header(hv):
                continue
            size_cols[normalizar_nombre(hv)] = (c, str(hv).strip())
        if size_cols:
            rr = r + 1
            while rr <= max_row:
                componente = v(rr, 1)
                if componente is None or str(componente).strip() == "":
                    rr += 1
                    continue
                componente_txt = str(componente).strip()
                componente_norm = normalizar_nombre(componente_txt)
                if componente_norm in {"subtotal 1", "costo sin m o"} or componente_norm.startswith("subtotal"):
                    break
                if componente_norm.startswith("costo"):
                    break

                for _, (col, presentacion) in size_cols.items():
                    costo = _to_number(v(rr, col))
                    if costo == "":
                        continue
                    ingrediente_txt = componente_txt
                    if componente_norm == "pan":
                        pan_name = _resolve_pan_from_formula(vf(rr, col), pan_map or {})
                        if pan_name:
                            ingrediente_txt = f"{pan_name} - {presentacion}"
                    receta_nombre = f"{ws.title} - {presentacion}"[:250]
                    orden_actual = orden_por_receta.get(receta_nombre, 0) + 1
                    orden_por_receta[receta_nombre] = orden_actual
                    rows.append(
                        {
                            "receta": receta_nombre,
                            "subreceta": ws.title[:120],
                            "producto_final": ws.title[:250],
                            "tipo": "PRODUCTO_FINAL",
                            "tipo_linea": "NORMAL",
                            "etapa": "",
                            "ingrediente": ingrediente_txt[:250],
                            "cantidad": "",
                            "unidad": "",
                            "costo_linea": costo,
                            "orden": orden_actual,
                            "notas": "",
                        }
                    )
                rr += 1
        break

    # Bloques de subsección: Elemento + tamaños (Cobertura/Relleno/etc dentro de Dream Whip, Fresa, etc.).
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if normalizar_nombre(v(r, c)) != "elemento":
                continue

            headers: list[tuple[int, str]] = []
            cc = c + 1
            while cc <= max_col:
                hv = v(r, cc)
                if hv is None or str(hv).strip() == "":
                    if headers:
                        break
                    cc += 1
                    continue
                if isinstance(hv, (int, float)):
                    break
                htxt = str(hv).strip()
                if _is_presentation_header(htxt):
                    headers.append((cc, htxt))
                cc += 1

            if not headers:
                continue

            section = ""
            if r > 1:
                prev = v(r - 1, c)
                if isinstance(prev, str) and prev.strip():
                    section = prev.strip()[:120]

            rr = r + 1
            while rr <= max_row:
                elemento = v(rr, c)
                if elemento is None or str(elemento).strip() == "":
                    break

                elemento_txt = str(elemento).strip()
                elemento_norm = normalizar_nombre(elemento_txt)
                if elemento_norm == "elemento" or elemento_norm.startswith("total"):
                    break

                for col, presentacion in headers:
                    cantidad = _to_number(v(rr, col))
                    if cantidad == "":
                        continue

                    receta_nombre = f"{ws.title} - {presentacion.strip()}"[:250]
                    orden_actual = orden_por_receta.get(receta_nombre, 0) + 1
                    orden_por_receta[receta_nombre] = orden_actual

                    rows.append(
                        {
                            "receta": receta_nombre,
                            "subreceta": section or ws.title[:120],
                            "producto_final": ws.title[:250],
                            "tipo": "PRODUCTO_FINAL",
                            "tipo_linea": "SUBSECCION",
                            "etapa": section,
                            "ingrediente": elemento_txt[:250],
                            "cantidad": cantidad,
                            "unidad": "kg",
                            "costo_linea": "",
                            "orden": orden_actual,
                            "notas": section,
                        }
                    )
                rr += 1

    return rows


def convert_costeo_to_template(costeo_path: str) -> tuple[list[dict[str, Any]], ConversionResult]:
    path = Path(costeo_path)
    if not path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {costeo_path}")

    wb = load_workbook(path, data_only=True)
    wb_formula = load_workbook(path, data_only=False)
    pan_map: dict[int, str] = {}
    if "Insumos 2" in wb_formula.sheetnames:
        pan_map = _build_insumos2_pan_map(wb_formula["Insumos 2"])
    all_rows: list[dict[str, Any]] = []
    seen_recetas: set[tuple[str, str]] = set()
    result = ConversionResult(hojas_escaneadas=len(wb.sheetnames))

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_formula = wb_formula[sheet_name] if sheet_name in wb_formula.sheetnames else ws
        rows = _find_recipe_block(ws)
        product_rows = _find_product_final_matrix(ws, ws_formula=ws_formula, pan_map=pan_map)
        merged_rows = rows + product_rows
        if merged_rows:
            result.hojas_con_recetas += 1
        for row in merged_rows:
            all_rows.append(row)
            seen_recetas.add((row["subreceta"], row["receta"]))

    result.recetas_detectadas = len(seen_recetas)
    result.lineas_detectadas = len(all_rows)
    return all_rows, result


def write_template_csv(rows: list[dict[str, Any]], output_path: str) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=TEMPLATE_HEADERS)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_template_xlsx(rows: list[dict[str, Any]], output_path: str) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "recetas"
    ws.append(TEMPLATE_HEADERS)
    for row in rows:
        ws.append([row.get(h, "") for h in TEMPLATE_HEADERS])
    wb.save(path)
