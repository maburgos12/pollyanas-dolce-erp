import hashlib
import csv
from io import BytesIO
from math import exp, sqrt
from datetime import date, datetime, timedelta
from calendar import monthrange
from collections import defaultdict
from decimal import Decimal, ROUND_CEILING
from typing import Dict, Any, List
from urllib.parse import urlencode
from uuid import uuid4

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.exceptions import PermissionDenied
from django.db import OperationalError, ProgrammingError, connection, transaction
from django.db.models.deletion import ProtectedError
from django.db.models import Count, Q, OuterRef, Subquery, Case, When, Value, IntegerField, Sum, DecimalField, Max, Exists
from django.core.paginator import Paginator
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.urls import reverse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from rapidfuzz import fuzz

from compras.models import OrdenCompra, RecepcionCompra, SolicitudCompra
from control.models import MermaPOS
from core.access import can_manage_compras, can_view_recetas, is_branch_capture_only
from core.audit import log_event
from core.models import Sucursal, sucursales_operativas
from inventario.models import ExistenciaInsumo, MovimientoInventario
from maestros.models import CostoInsumo, Insumo, UnidadMedida
from pos_bridge.config import load_point_bridge_settings
from pos_bridge.models import (
    PointDailyBranchIndicator,
    PointDailySale,
    PointInventorySnapshot,
    PointProduct,
    PointSalesDailyProductFact,
    PointSyncJob,
    PointProductionLine,
    PointTransferLine,
    PointWasteLine,
    PointProductCategory,
)
from pos_bridge.services.product_recipe_sync_service import PointProductRecipeSyncService
from maestros.utils.canonical_catalog import (
    canonical_insumo,
    canonical_insumo_by_id,
    canonicalized_insumo_selector,
    latest_costo_canonico,
)
from ..models import (
    Receta,
    RecetaAgrupacionAddon,
    RecetaCodigoPointAlias,
    RecetaCostoSemanal,
    RecetaEquivalencia,
    normalizar_codigo_point,
    LineaReceta,
    RecetaPresentacion,
    RecetaPresentacionDerivada,
    RecetaCostoVersion,
    CostoDriver,
    PlanProduccion,
    PlanProduccionItem,
    PronosticoVenta,
    SolicitudVenta,
    VentaHistorica,
    PoliticaStockSucursalProducto,
    InventarioCedisProducto,
    MovimientoProductoCedis,
    SolicitudReabastoCedis,
    SolicitudReabastoCedisLinea,
)
from ..utils.costeo_versionado import asegurar_version_costeo, calcular_costeo_receta, comparativo_versiones
from ..utils.costeo_semanal import snapshot_weekly_costs, week_bounds
from ..utils.costeo_snapshot import resolve_insumo_unit_cost, resolve_line_snapshot_cost
from ..utils.derived_product_presentations import (
    build_upstream_snapshot as build_derived_product_upstream_snapshot,
    get_active_derived_relation,
)
from ..utils.derived_insumos import sync_presentacion_insumo, sync_receta_derivados
from ..utils.matching import match_insumo
from ..utils.normalizacion import normalizar_nombre
from ..catalogs import familia_categoria_catalogo_json, familias_producto_catalogo
from reportes.executive_panels import _partial_month_amount_quantity
from ventas.models import VentaAutoritativaPoint
from ventas.services.financials import resolve_unit_prices_bulk

OFFICIAL_POINT_SOURCE = "/Report/PrintReportes?idreporte=3"
RECENT_POINT_SOURCE = "/Report/VentasCategorias"
NONBLOCKING_PRODUCT_PLACEHOLDER_LABELS = {
    "armado",
    "presentacion",
}
DIRECT_BASE_REPLACEMENT_MIN_QTY = Decimal("0.25")
DIRECT_BASE_REPLACEMENT_MAX_QTY = Decimal("4.00")
FORECAST_HISTORY_START = date(2022, 1, 1)
FORECAST_RECENCY_HALF_LIFE_DAYS = Decimal("540")
FORECAST_MIN_RECENCY_WEIGHT = Decimal("0.20")
FORECAST_MARCH_2026_LEGACY_WEIGHT = Decimal("0.20")
FORECAST_EXCLUDED_MONTHS = {(2026, 3)}
MIX_ADJUSTMENT_ENABLED_DEFAULT = False
MIX_RECENT_WINDOW_DAYS = 84
MIX_SHORT_WINDOW_DAYS = 28
MIX_MIN_RECENT_SALES_QTY = Decimal("12")
MIX_MIN_EFFECTIVE_RECENT_SAMPLES = Decimal("4")
MIX_MAX_CHANGE_PCT = Decimal("0.20")
MIX_ALPHA_MAX = Decimal("0.35")
RECENT_PRODUCT_ACTIVITY_DAYS = 90


def _presentacion_sort_key(nombre: str) -> tuple[int, str]:
    norm = normalizar_nombre(nombre or "")
    order = {
        "mini": 10,
        "chico": 20,
        "mediano": 30,
        "grande": 40,
        "bollos": 50,
        "bollo": 50,
        "bollito": 50,
        "individual": 60,
        "media plancha": 70,
        "1 2 plancha": 70,
        "rosca": 80,
    }
    for token, rank in order.items():
        if f" {token}" in f" {norm}" or norm.endswith(token):
            return rank, norm
    return 999, norm


def _insumo_canonical_priority(insumo: Insumo) -> int:
    score = 0
    if (insumo.codigo_point or "").strip():
        score += 100
    latest_cost = getattr(insumo, "latest_costo_unitario", None)
    if latest_cost is not None and Decimal(str(latest_cost or 0)) > 0:
        score += 60
    if insumo.proveedor_principal_id:
        score += 30
    if insumo.unidad_base_id:
        score += 20
    if (insumo.codigo or "").strip():
        score += 10
    return score


def _canonicalize_insumo_match(insumo: Insumo | None) -> Insumo | None:
    return canonical_insumo(insumo)


def _insumo_erp_readiness(insumo: Insumo) -> dict[str, object]:
    missing: list[str] = []
    codigo_norm = (insumo.codigo or "").strip().upper()
    is_recipe_derived = codigo_norm.startswith("DERIVADO:RECETA:")
    effective_tipo = insumo.tipo_item
    if is_recipe_derived:
        effective_tipo = Insumo.TIPO_INTERNO
    if not insumo.activo:
        missing.append("inactivo")
    if not insumo.unidad_base_id:
        missing.append("unidad base")
    if effective_tipo == Insumo.TIPO_MATERIA_PRIMA and not insumo.proveedor_principal_id:
        missing.append("proveedor principal")
    if (
        effective_tipo == Insumo.TIPO_INTERNO
        and not is_recipe_derived
        and not (insumo.categoria or "").strip()
    ):
        missing.append("categoría")
    if effective_tipo == Insumo.TIPO_EMPAQUE and not (insumo.categoria or "").strip():
        missing.append("categoría")
    return {
        "ready": not missing,
        "missing": missing,
        "label": "Listo para operar" if not missing else "Incompleto",
    }


def _insumo_operational_readiness(
    insumo: Insumo,
    *,
    ignore_supplier: bool = False,
) -> dict[str, object]:
    profile = _insumo_erp_readiness(insumo)
    missing = list(profile["missing"])
    if ignore_supplier:
        missing = [item for item in missing if item != "proveedor principal"]
    return {
        "ready": not missing,
        "missing": missing,
        "label": "Listo para operar" if not missing else "Incompleto",
    }


def _insumo_display_name(insumo: Insumo | None) -> str:
    if not insumo:
        return ""
    return (insumo.nombre_point or insumo.nombre or "").strip()


def _insumo_article_class(insumo: Insumo) -> dict[str, str]:
    if insumo.tipo_item == Insumo.TIPO_EMPAQUE:
        return {
            "key": Insumo.TIPO_EMPAQUE,
            "label": "Empaque",
        }
    if insumo.tipo_item == Insumo.TIPO_INTERNO or (insumo.codigo or "").startswith("DERIVADO:RECETA:"):
        return {
            "key": Insumo.TIPO_INTERNO,
            "label": "Insumo interno",
        }
    return {
        "key": Insumo.TIPO_MATERIA_PRIMA,
        "label": "Materia prima",
    }


def _filtered_recipe_sync_queryset(params) -> Any:
    vista = (params.get("vista") or "").strip().lower()
    q = (params.get("q") or "").strip()
    tipo = (params.get("tipo") or "").strip().upper()
    modo_operativo = (params.get("modo_operativo") or "").strip().upper()
    familia = (params.get("familia") or "").strip()
    categoria = (params.get("categoria") or "").strip()

    qs = Receta.objects.all().order_by("nombre", "id")
    if vista == "productos":
        qs = qs.filter(tipo=Receta.TIPO_PRODUCTO_FINAL)
    elif vista == "insumos":
        qs = qs.filter(tipo=Receta.TIPO_PREPARACION)
    elif vista == "subinsumos":
        qs = qs.filter(tipo=Receta.TIPO_PREPARACION, usa_presentaciones=True)

    if tipo in {Receta.TIPO_PREPARACION, Receta.TIPO_PRODUCTO_FINAL}:
        qs = qs.filter(tipo=tipo)
    if modo_operativo == "BASE":
        qs = qs.filter(tipo=Receta.TIPO_PREPARACION, usa_presentaciones=False)
    elif modo_operativo == "BASE_DERIVADOS":
        qs = qs.filter(tipo=Receta.TIPO_PREPARACION, usa_presentaciones=True)
    elif modo_operativo == "FINAL":
        qs = qs.filter(tipo=Receta.TIPO_PRODUCTO_FINAL)
    if familia:
        qs = qs.filter(familia=familia)
    if categoria:
        qs = qs.filter(categoria=categoria)
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(codigo_point__icontains=q))
    return qs


def _run_point_recipe_sync_action(
    request: HttpRequest,
    *,
    action_label: str,
    product_codes: list[str] | None = None,
    branch_hint: str = "MATRIZ",
    include_without_recipe: bool = False,
) -> tuple[PointSyncJob, dict[str, Any]]:
    clean_codes = [str(code).strip().upper() for code in (product_codes or []) if str(code).strip()]
    job = PointSyncJob.objects.create(
        job_type=PointSyncJob.JOB_TYPE_RECIPES,
        status=PointSyncJob.STATUS_RUNNING,
        parameters={
            "action": action_label,
            "branch_hint": branch_hint,
            "product_codes": clean_codes,
            "include_without_recipe": include_without_recipe,
        },
        triggered_by=request.user,
    )
    service = PointProductRecipeSyncService()
    try:
        result = service.sync(
            branch_hint=branch_hint,
            product_codes=clean_codes or None,
            include_without_recipe=include_without_recipe,
            sync_job=job,
        )
        job.status = PointSyncJob.STATUS_SUCCESS
        job.finished_at = timezone.now()
        job.result_summary = result.summary
        job.artifacts = {"raw_export_path": result.raw_export_path}
        job.save(update_fields=["status", "finished_at", "result_summary", "artifacts", "updated_at"])
        request.session["recetas_last_point_sync_job_id"] = job.id
        log_event(
            request.user,
            "SYNC_POINT_RECIPES",
            "pos_bridge.PointSyncJob",
            job.id,
            {
                "action": action_label,
                "product_codes": clean_codes,
                "summary": result.summary,
                "raw_export_path": result.raw_export_path,
            },
        )
        return job, result.summary
    except Exception as exc:
        job.status = PointSyncJob.STATUS_FAILED
        job.finished_at = timezone.now()
        job.error_message = str(exc)
        job.save(update_fields=["status", "finished_at", "error_message", "updated_at"])
        log_event(
            request.user,
            "SYNC_POINT_RECIPES_FAILED",
            "pos_bridge.PointSyncJob",
            job.id,
            {
                "action": action_label,
                "product_codes": clean_codes,
                "error": str(exc),
            },
        )
        raise


def _format_point_recipe_sync_message(summary: dict[str, Any], *, new_codes_count: int | None = None) -> str:
    completed = int(summary.get("recipes_completed_successfully") or 0)
    warnings = int(summary.get("recipes_with_unresolved_inputs") or 0)
    new_products = int(summary.get("new_products_imported") or 0)
    new_preparations = int(summary.get("new_preparations_imported") or 0)
    unresolved = int(summary.get("unresolved_inputs_count") or 0)
    products_selected = int(summary.get("products_selected") or 0)
    products_label = new_codes_count if new_codes_count is not None else products_selected

    parts = [
        f"Point procesó {products_selected} producto(s)",
        f"{completed} completo(s)",
        f"{warnings} con advertencias",
    ]
    if new_codes_count is not None:
        parts.insert(0, f"Se detectaron {products_label} código(s) nuevo(s)")
    if new_products:
        parts.append(f"{new_products} producto(s) nuevo(s) creado(s)")
    if new_preparations:
        parts.append(f"{new_preparations} preparación(es) hija(s) nueva(s)")
    if unresolved:
        parts.append(f"{unresolved} insumo(s) pendiente(s)")
    return ". ".join(parts) + "."


def _format_point_recipe_discovery_blocked_message(discovery: dict[str, Any]) -> str:
    blocked_candidates = discovery.get("blocked_candidates") or []
    blocked_count = int(discovery.get("blocked_candidates_count") or len(blocked_candidates))
    if blocked_count <= 0:
        return "Point no reportó productos nuevos con receta pendientes de incorporar."
    sample = ", ".join(
        f"{item.get('codigo_point') or ''} {item.get('nombre') or ''}".strip()
        for item in blocked_candidates[:3]
    )
    suffix = f" Ejemplos: {sample}." if sample else ""
    return (
        f"Point no reportó productos nuevos importables con receta, pero detectó {blocked_count} "
        f"candidato(s) nuevo(s) bloqueado(s) porque Point no entregó receta/BOM.{suffix}"
    )


def _point_recipe_sync_job_panel(request: HttpRequest) -> dict[str, Any] | None:
    job_qs = PointSyncJob.objects.filter(job_type=PointSyncJob.JOB_TYPE_RECIPES)
    preferred_job_id = request.session.get("recetas_last_point_sync_job_id")
    job = None
    if preferred_job_id:
        candidate = job_qs.filter(id=preferred_job_id).first()
        if candidate is not None and (candidate.parameters or {}).get("mode") != "recipe_gap_audit":
            job = candidate
    if job is None:
        for candidate in job_qs:
            if (candidate.parameters or {}).get("mode") != "recipe_gap_audit":
                job = candidate
                break
    if job is None:
        return None

    summary = job.result_summary or {}
    imported_products = list(summary.get("imported_products_status") or [])
    status_label_map = {
        "SUCCESS_COMPLETE": ("Completo", "success"),
        "SUCCESS_WITH_WARNINGS": ("Con advertencias", "warning"),
        "BLOCKED_UNRESOLVED": ("Bloqueado", "danger"),
    }
    decorated_products: list[dict[str, Any]] = []
    for item in imported_products[:5]:
        label, tone = status_label_map.get(item.get("status"), ("Sin clasificar", "warning"))
        decorated_products.append(
            {
                **item,
                "status_label": label,
                "tone": tone,
            }
        )

    return {
        "job_id": job.id,
        "job_status": job.status,
        "job_status_label": job.get_status_display(),
        "started_at": job.started_at,
        "finished_at": job.finished_at,
        "summary": summary,
        "products": decorated_products,
        "products_more_count": max(0, len(imported_products) - len(decorated_products)),
    }


def _snapshot_costeo_after_sync(*, receta_ids: list[int] | None = None) -> dict[str, Any]:
    summary = snapshot_weekly_costs(receta_ids=receta_ids)
    return {
        "week_start": summary.week_start.isoformat(),
        "week_end": summary.week_end.isoformat(),
        "recipes_created": summary.recipes_created,
        "recipes_updated": summary.recipes_updated,
        "addons_created": summary.addons_created,
        "addons_updated": summary.addons_updated,
        "total_items": summary.total_items,
    }


def _missing_field_to_filter_key(missing_label: str) -> str | None:
    mapping = {
        "unidad base": "unidad",
        "proveedor principal": "proveedor",
        "categoría": "categoria",
        "código comercial": "codigo_point",
        "codigo comercial": "codigo_point",
        "código externo": "codigo_point",
        "codigo externo": "codigo_point",
    }
    return mapping.get(missing_label)


def _enterprise_blocker_action_meta_for_recipes(
    item_name: str,
    class_key: str,
    missing_field: str | None,
    insumo_id: int | None = None,
    usage_scope: str = "recipes",
) -> dict[str, str]:
    missing_value = (missing_field or "").strip()
    normalized = normalizar_nombre(missing_value)
    query = {
        "tipo_item": class_key,
        "enterprise_status": "incompletos",
        "usage_scope": usage_scope,
    }
    filter_key = _missing_field_to_filter_key(missing_value) if missing_value else None
    if filter_key:
        query["missing_field"] = filter_key
    if insumo_id:
        query["insumo_id"] = int(insumo_id)
    elif item_name:
        query["q"] = item_name
    list_url = reverse("maestros:insumo_list") + f"?{urlencode(query)}"
    edit_url = reverse("maestros:insumo_update", args=[insumo_id]) if insumo_id else ""

    if "unidad" in normalized:
        return {
            "label": "Definir unidad base",
            "detail": "Asigna la unidad base para que costeo, recetas y abastecimiento trabajen con la misma unidad operativa.",
            "url": list_url,
            "edit_url": edit_url,
        }
    if "proveedor" in normalized:
        return {
            "label": "Asignar proveedor principal",
            "detail": "Define el proveedor principal para habilitar compras, lead time y reposición del artículo.",
            "url": list_url,
            "edit_url": edit_url,
        }
    if "categoria" in normalized:
        return {
            "label": "Asignar categoría",
            "detail": "Completa la categoría para clasificar el artículo correctamente en recetas, compras y reportes.",
            "url": list_url,
            "edit_url": edit_url,
        }
    if "codigo point" in normalized or "point" in normalized or "codigo comercial" in normalized or "código comercial" in normalized:
        return {
            "label": "Registrar código comercial",
            "detail": "Registra el código comercial para conciliar ventas, surtido y catálogo operativo.",
            "url": list_url,
            "edit_url": edit_url,
        }
    if "costo" in normalized:
        return {
            "label": "Cargar costo vigente",
            "detail": "Registra un costo vigente para que costeo, MRP y compras calculen valores reales.",
            "url": list_url,
            "edit_url": edit_url,
        }
    if "inactivo" in normalized:
        return {
            "label": "Reactivar artículo",
            "detail": "Reactiva el artículo o sustitúyelo por su versión maestra antes de usarlo en recetas o abastecimiento.",
            "url": list_url,
            "edit_url": edit_url,
        }
    if "catalogo" in normalized or "canonico" in normalized:
        return {
            "label": "Revisar catálogo",
            "detail": "Consolida variantes y usa el artículo estándar para evitar duplicidad en costeo, compras y MRP.",
            "url": list_url,
            "edit_url": edit_url,
        }
    return {
        "label": "Abrir maestro",
        "detail": "Completa el maestro del artículo para liberar costeo, MRP, recetas y compras.",
        "url": list_url,
        "edit_url": edit_url,
    }


def _enterprise_blocker_label_detail_for_missing_recipes(missing_field: str | None) -> tuple[str, str]:
    meta = _enterprise_blocker_action_meta_for_recipes("", "", missing_field, usage_scope="recipes")
    return meta["label"], meta["detail"]


def _attach_linea_suggested_match(linea: LineaReceta) -> None:
    linea.suggested_insumo = None
    linea.suggested_score = 0.0
    linea.suggested_method = "NO_MATCH"
    linea.suggested_can_approve = False
    linea.suggested_is_canonical = False
    if linea.tipo_linea == LineaReceta.TIPO_SUBSECCION or linea.insumo_id:
        return
    if linea.match_status not in {LineaReceta.STATUS_NEEDS_REVIEW, LineaReceta.STATUS_REJECTED}:
        return
    if not (linea.insumo_texto or "").strip():
        return

    raw_matched, raw_score, raw_method = match_insumo(linea.insumo_texto)
    suggested_insumo = _canonicalize_insumo_match(raw_matched)
    linea.suggested_insumo = suggested_insumo
    linea.suggested_score = float(raw_score or 0.0)
    linea.suggested_method = raw_method
    linea.suggested_can_approve = bool(suggested_insumo and float(raw_score or 0.0) >= 75.0)
    linea.suggested_is_canonical = bool(
        suggested_insumo and raw_matched and suggested_insumo.id != raw_matched.id
    )


def _attach_linea_canonical_target(linea: LineaReceta) -> None:
    linea.canonical_target = None
    linea.canonical_needs_repoint = False
    if not linea.insumo_id:
        return
    canonical = _canonicalize_insumo_match(linea.insumo)
    if canonical and canonical.id != linea.insumo_id:
        linea.canonical_target = canonical
        linea.canonical_needs_repoint = True


def _validate_selected_insumo_enterprise_ready(linea: LineaReceta) -> str | None:
    if not linea.insumo_id:
        return None
    profile = _insumo_erp_readiness(linea.insumo)
    blocking_missing = [item for item in profile["missing"] if item != "proveedor principal"]
    if not blocking_missing:
        return None
    if blocking_missing == ["inactivo"]:
        return f"El artículo seleccionado ({linea.insumo.nombre}) está inactivo. Usa un artículo vigente."
    missing_human = ", ".join(blocking_missing)
    return (
        f"El artículo seleccionado ({linea.insumo.nombre}) no está listo para operar en ERP. "
        f"Corrige en Maestros: {missing_human}."
    )


def _recipe_search_score(query_norm: str, receta: Receta) -> float:
    """Score de coincidencia aproximada para búsqueda en listado de recetas."""
    if not query_norm:
        return 0.0

    fields = [
        receta.nombre or "",
        receta.codigo_point or "",
        receta.familia or "",
        receta.categoria or "",
        receta.sheet_name or "",
    ]

    best = 0.0
    for field in fields:
        field_norm = normalizar_nombre(field)
        if not field_norm:
            continue
        if query_norm == field_norm:
            return 100.0
        if query_norm in field_norm:
            best = max(best, 96.0)
            continue
        score = max(
            float(fuzz.WRatio(query_norm, field_norm)),
            float(fuzz.partial_ratio(query_norm, field_norm)),
            float(fuzz.token_set_ratio(query_norm, field_norm)),
        )
        best = max(best, score)
    return best


def _recipe_derived_sync_state(receta: Receta) -> dict[str, int | bool]:
    cached = getattr(receta, "_derived_sync_state_cache", None)
    if cached is not None:
        return cached

    prep_code = f"DERIVADO:RECETA:{receta.id}:PREPARACION"
    presentacion_prefix = f"DERIVADO:RECETA:{receta.id}:PRESENTACION:"
    presentaciones_activas = getattr(receta, "presentaciones_activas_count", None)
    if presentaciones_activas is None:
        presentaciones_activas = receta.presentaciones.filter(activo=True).count()

    state = {
        "prep_ready": Insumo.objects.filter(codigo=prep_code, activo=True).exists(),
        "active_presentaciones": int(presentaciones_activas or 0),
        "derived_presentaciones": Insumo.objects.filter(
            codigo__startswith=presentacion_prefix,
            activo=True,
        ).count(),
    }
    setattr(receta, "_derived_sync_state_cache", state)
    return state


def _recipe_supply_chain_snapshot(receta: Receta) -> dict[str, object] | None:
    if receta.tipo != Receta.TIPO_PREPARACION:
        return None

    cached = getattr(receta, "_supply_chain_snapshot_cache", None)
    if cached is not None:
        return cached

    prep_code = f"DERIVADO:RECETA:{receta.id}:PREPARACION"
    presentacion_prefix = f"DERIVADO:RECETA:{receta.id}:PRESENTACION:"

    prep_insumo = (
        Insumo.objects.filter(codigo=prep_code, activo=True)
        .only("id", "nombre", "codigo")
        .first()
    )
    presentacion_qs = Insumo.objects.filter(
        codigo__startswith=presentacion_prefix,
        activo=True,
    ).only("id", "nombre", "codigo")
    presentacion_items = [
        {"id": item.id, "nombre": item.nombre}
        for item in presentacion_qs.order_by("nombre")
    ]
    presentacion_ids = list(presentacion_qs.values_list("id", flat=True))
    component_ids = []
    if prep_insumo:
        component_ids.append(prep_insumo.id)
    component_ids.extend(presentacion_ids)

    downstream_line_qs = LineaReceta.objects.filter(
        receta__tipo=Receta.TIPO_PRODUCTO_FINAL,
        insumo_id__in=component_ids,
    ).select_related("receta", "insumo")
    downstream_recipes = list(
        downstream_line_qs.order_by("receta__nombre")
        .values("receta_id", "receta__nombre")
        .distinct()
    )
    snapshot = {
        "prep_insumo": prep_insumo,
        "prep_insumo_item": (
            {"id": prep_insumo.id, "nombre": prep_insumo.nombre}
            if prep_insumo
            else None
        ),
        "presentacion_count": len(presentacion_ids),
        "presentacion_items": presentacion_items[:8],
        "component_count": len(component_ids),
        "downstream_line_count": downstream_line_qs.count(),
        "downstream_recipe_count": len(downstream_recipes),
        "downstream_recipes": [
            {"id": item["receta_id"], "nombre": item["receta__nombre"]}
            for item in downstream_recipes[:6]
        ],
        "has_downstream_usage": bool(downstream_recipes),
    }
    setattr(receta, "_supply_chain_snapshot_cache", snapshot)
    return snapshot


def _recipe_id_from_derived_code(codigo: str) -> int | None:
    raw = (codigo or "").strip()
    if not raw.startswith("DERIVADO:RECETA:"):
        return None
    parts = raw.split(":")
    if len(parts) < 3:
        return None
    try:
        return int(parts[2])
    except (TypeError, ValueError):
        return None


def _derived_code_kind(codigo: str) -> str | None:
    raw = (codigo or "").strip()
    if not raw.startswith("DERIVADO:RECETA:"):
        return None
    parts = raw.split(":")
    if len(parts) < 4:
        return None
    if parts[3] == "PREPARACION":
        return "PREPARACION"
    if parts[3] == "PRESENTACION":
        return "PRESENTACION"
    return None


def _product_upstream_snapshot(lineas: list[LineaReceta], *, receta: Receta | None = None) -> dict[str, object]:
    source_map: dict[int, dict[str, object]] = {}
    internal_count = 0
    internal_without_source_count = 0
    empaque_count = 0
    mp_count = 0
    for linea in lineas:
        if not linea.insumo_id:
            continue
        if linea.insumo.tipo_item == Insumo.TIPO_EMPAQUE:
            empaque_count += 1
        elif linea.insumo.tipo_item == Insumo.TIPO_MATERIA_PRIMA:
            mp_count += 1
        else:
            internal_count += 1

        source_recipe_id = _recipe_id_from_derived_code(linea.insumo.codigo or "")
        if not source_recipe_id:
            if linea.insumo.tipo_item != Insumo.TIPO_EMPAQUE and linea.insumo.tipo_item != Insumo.TIPO_MATERIA_PRIMA:
                internal_without_source_count += 1
            continue
        current = source_map.setdefault(
            source_recipe_id,
            {
                "id": source_recipe_id,
                "nombre": getattr(getattr(linea, "source_recipe", None), "nombre", ""),
                "line_count": 0,
            },
        )
        if not current["nombre"] and getattr(linea, "source_recipe", None):
            current["nombre"] = linea.source_recipe.nombre
        current["line_count"] += 1

    missing_name_ids = [item["id"] for item in source_map.values() if not item["nombre"]]
    if missing_name_ids:
        for recipe in Receta.objects.filter(id__in=missing_name_ids).only("id", "nombre"):
            if recipe.id in source_map:
                source_map[recipe.id]["nombre"] = recipe.nombre

    derived_parent_snapshot = build_derived_product_upstream_snapshot(receta) if receta is not None else None
    if derived_parent_snapshot:
        source_map.setdefault(
            int(derived_parent_snapshot["parent_recipe_id"]),
            {
                "id": int(derived_parent_snapshot["parent_recipe_id"]),
                "nombre": str(derived_parent_snapshot["parent_recipe_name"]),
                "line_count": 1,
                "derived_parent_link": True,
            },
        )

    upstream_bases = sorted(source_map.values(), key=lambda item: ((item["nombre"] or "").lower(), item["id"]))
    return {
        "upstream_bases": upstream_bases[:8],
        "upstream_base_count": len(upstream_bases),
        "internal_count": internal_count,
        "internal_without_source_count": internal_without_source_count,
        "internal_with_source_count": max(internal_count - internal_without_source_count, 0),
        "empaque_count": empaque_count,
        "mp_count": mp_count,
        "derived_parent_snapshot": derived_parent_snapshot,
    }


def _recipe_requires_fixed_packaging(receta: Receta) -> bool:
    if receta.tipo != Receta.TIPO_PRODUCTO_FINAL:
        return False
    nombre_norm = normalizar_nombre(receta.nombre or "")
    familia_norm = normalizar_nombre(receta.familia or "")
    categoria_norm = normalizar_nombre(receta.categoria or "")
    # Reglas operativas confirmadas por DG:
    # - Bollos y galletas no se entregan con empaque unitario fijo.
    # - Su salida depende de empaque por paquete o decisión comercial al momento.
    # - Empanadas individuales no llevan empaque unitario fijo.
    # - Bolitas de Nuez KG se maneja sin empaque unitario directo por ahora.
    if nombre_norm.startswith("bollo ") or nombre_norm.startswith("galleta "):
        return False
    if familia_norm == "empanadas" or categoria_norm == "empanadas" or nombre_norm.startswith("empanada "):
        return False
    if (receta.codigo_point or "").strip() == "05021" or nombre_norm == "bolitas de nuez kg":
        return False
    return True


def _recipe_packaging_ready(receta: Receta, upstream_snapshot: dict[str, object]) -> bool:
    if not _recipe_requires_fixed_packaging(receta):
        return True
    return int(upstream_snapshot.get("empaque_count") or 0) > 0


def _recipe_packaging_missing(receta: Receta, upstream_snapshot: dict[str, object]) -> bool:
    return _recipe_requires_fixed_packaging(receta) and int(upstream_snapshot.get("empaque_count") or 0) <= 0


def _direct_base_usage_snapshot(lineas: list[LineaReceta]) -> dict[str, object]:
    source_recipe_ids = {
        recipe_id
        for linea in lineas
        if getattr(linea, "insumo_id", None)
        for recipe_id in [_recipe_id_from_derived_code(getattr(getattr(linea, "insumo", None), "codigo", "") or "")]
        if recipe_id
    }
    if not source_recipe_ids:
        return {"count": 0, "base_names": []}

    source_recipe_map = {
        item.id: item
        for item in Receta.objects.filter(id__in=source_recipe_ids).only("id", "nombre", "usa_presentaciones")
    }
    source_recipe_presentaciones = {
        row["receta_id"]: int(row["total"])
        for row in (
            RecetaPresentacion.objects.filter(receta_id__in=source_recipe_ids, activo=True)
            .values("receta_id")
            .annotate(total=Count("id"))
        )
    }

    direct_base_names: list[str] = []
    count = 0
    replacement_cache: dict[int, list[dict[str, object]]] = {}
    for linea in lineas:
        insumo = getattr(linea, "insumo", None)
        if not getattr(linea, "insumo_id", None) or not insumo or insumo.tipo_item != Insumo.TIPO_INTERNO:
            continue
        if _derived_code_kind(insumo.codigo or "") != "PREPARACION":
            continue
        source_recipe_id = _recipe_id_from_derived_code(insumo.codigo or "")
        source_recipe = source_recipe_map.get(source_recipe_id) if source_recipe_id else None
        active_presentaciones = int(source_recipe_presentaciones.get(source_recipe_id, 0)) if source_recipe_id else 0
        if not source_recipe or not source_recipe.usa_presentaciones or active_presentaciones <= 0:
            continue
        candidates = _active_presentation_derived_candidates(source_recipe_id, cache=replacement_cache)
        replacement = _suggest_direct_base_replacement(linea, cache=replacement_cache)
        if candidates and not replacement:
            continue
        count += 1
        if source_recipe.nombre not in direct_base_names:
            direct_base_names.append(source_recipe.nombre)

    return {
        "count": count,
        "base_names": direct_base_names[:4],
    }


def _recipe_direct_base_snapshot(receta: Receta) -> dict[str, object]:
    cached = getattr(receta, "_direct_base_snapshot_cache", None)
    if cached is not None:
        return cached

    if receta.tipo != Receta.TIPO_PRODUCTO_FINAL:
        snapshot = {
            "count": 0,
            "base_names": [],
            "suggested_count": 0,
            "exact_count": 0,
            "sample_suggestions": [],
        }
        setattr(receta, "_direct_base_snapshot_cache", snapshot)
        return snapshot

    lineas = list(
        receta.lineas.select_related("insumo").only(
            "id",
            "cantidad",
            "insumo_id",
            "insumo__id",
            "insumo__nombre",
            "insumo__codigo",
            "insumo__tipo_item",
        )
    )
    replacement_cache: dict[int, list[dict[str, object]]] = {}
    source_recipe_ids = {
        recipe_id
        for linea in lineas
        if getattr(linea, "insumo_id", None)
        for recipe_id in [_recipe_id_from_derived_code(getattr(getattr(linea, "insumo", None), "codigo", "") or "")]
        if recipe_id
    }
    source_recipe_map = {
        item.id: item
        for item in Receta.objects.filter(id__in=source_recipe_ids).only("id", "nombre", "usa_presentaciones")
    }
    source_recipe_presentaciones = {
        row["receta_id"]: int(row["total"])
        for row in (
            RecetaPresentacion.objects.filter(receta_id__in=source_recipe_ids, activo=True)
            .values("receta_id")
            .annotate(total=Count("id"))
        )
    }

    suggested_count = 0
    exact_count = 0
    sample_suggestions: list[dict[str, object]] = []
    base_names: list[str] = []
    count = 0

    for linea in lineas:
        insumo = getattr(linea, "insumo", None)
        if not getattr(linea, "insumo_id", None) or not insumo or insumo.tipo_item != Insumo.TIPO_INTERNO:
            continue
        if _derived_code_kind(insumo.codigo or "") != "PREPARACION":
            continue
        source_recipe_id = _recipe_id_from_derived_code(insumo.codigo or "")
        source_recipe = source_recipe_map.get(source_recipe_id) if source_recipe_id else None
        active_presentaciones = int(source_recipe_presentaciones.get(source_recipe_id, 0)) if source_recipe_id else 0
        if not source_recipe or not source_recipe.usa_presentaciones or active_presentaciones <= 0:
            continue

        candidates = _active_presentation_derived_candidates(source_recipe_id, cache=replacement_cache)
        replacement = _suggest_direct_base_replacement(linea, cache=replacement_cache)
        if candidates and not replacement:
            continue
        count += 1
        if source_recipe.nombre not in base_names:
            base_names.append(source_recipe.nombre)

        linea.uses_direct_base_in_final = True
        if not replacement:
            continue
        suggested_count += 1
        if replacement["exact_match"]:
            exact_count += 1
        if len(sample_suggestions) < 3:
            sample_suggestions.append(
                {
                    "linea_id": linea.id,
                    "insumo": replacement["insumo"],
                    "presentacion": replacement["presentacion"],
                    "replacement_quantity": replacement["replacement_quantity"],
                    "reason": replacement["reason"],
                    "exact_match": replacement["exact_match"],
                }
            )

    snapshot = {
        "count": count,
        "base_names": base_names[:4],
        "suggested_count": suggested_count,
        "exact_count": exact_count,
        "sample_suggestions": sample_suggestions,
    }
    setattr(receta, "_direct_base_snapshot_cache", snapshot)
    return snapshot


def _active_presentation_derived_candidates(
    receta_base_id: int,
    cache: dict[int, list[dict[str, object]]] | None = None,
) -> list[dict[str, object]]:
    if cache is not None and receta_base_id in cache:
        return cache[receta_base_id]

    presentaciones = list(
        RecetaPresentacion.objects.filter(receta_id=receta_base_id, activo=True).order_by("nombre")
    )
    if not presentaciones:
        if cache is not None:
            cache[receta_base_id] = []
        return []

    code_prefix = f"DERIVADO:RECETA:{receta_base_id}:PRESENTACION:"
    insumos = list(
        Insumo.objects.filter(codigo__startswith=code_prefix, activo=True).only(
            "id",
            "nombre",
            "codigo",
            "unidad_base_id",
        )
    )
    exact_code_map = {item.codigo: item for item in insumos if (item.codigo or "").strip()}
    normalized_name_map = {
        normalizar_nombre(item.nombre or ""): item
        for item in insumos
        if (item.nombre or "").strip()
    }

    candidates: list[dict[str, object]] = []
    for presentacion in presentaciones:
        expected_code = f"DERIVADO:RECETA:{receta_base_id}:PRESENTACION:{presentacion.id}"
        derived_insumo = exact_code_map.get(expected_code)
        if derived_insumo is None:
            derived_insumo = normalized_name_map.get(
                normalizar_nombre(f"{presentacion.receta.nombre} - {presentacion.nombre}")
            )
        if derived_insumo is None:
            continue
        candidates.append(
            {
                "presentacion": presentacion,
                "insumo": derived_insumo,
                "peso": Decimal(str(presentacion.peso_por_unidad_kg or 0)),
            }
        )

    candidates.sort(
        key=lambda item: (
            _presentacion_sort_key(item["presentacion"].nombre),
            item["presentacion"].id,
        )
    )
    if cache is not None:
        cache[receta_base_id] = candidates
    return candidates


def _suggest_direct_base_replacement(
    linea: LineaReceta,
    cache: dict[int, list[dict[str, object]]] | None = None,
) -> dict[str, object] | None:
    if not getattr(linea, "insumo_id", None):
        return None
    insumo = getattr(linea, "insumo", None)
    if not insumo or insumo.tipo_item != Insumo.TIPO_INTERNO:
        return None
    if _derived_code_kind(getattr(insumo, "codigo", "") or "") != "PREPARACION":
        return None
    source_recipe_id = _recipe_id_from_derived_code(getattr(linea.insumo, "codigo", "") or "")
    if not source_recipe_id:
        return None
    qty = Decimal(str(getattr(linea, "cantidad", None) or 0))
    if qty <= 0:
        return None

    candidates = _active_presentation_derived_candidates(source_recipe_id, cache=cache)
    if not candidates:
        return None

    tolerance = Decimal("0.000001")
    best = min(
        candidates,
        key=lambda item: (
            abs(item["peso"] - qty),
            _presentacion_sort_key(item["presentacion"].nombre),
            item["presentacion"].id,
        ),
    )
    difference = abs(best["peso"] - qty)
    exact_match = difference <= tolerance
    replacement_quantity = Decimal("0")
    if best["peso"] > 0:
        replacement_quantity = (qty / best["peso"]).quantize(Decimal("0.000001"))
    if not exact_match and (
        replacement_quantity < DIRECT_BASE_REPLACEMENT_MIN_QTY
        or replacement_quantity > DIRECT_BASE_REPLACEMENT_MAX_QTY
    ):
        return None
    qty_human = f"{qty.quantize(Decimal('0.01'))}"
    peso_human = f"{best['peso'].quantize(Decimal('0.01'))}"
    if exact_match:
        reason = f"Coincide con la cantidad capturada ({qty_human} kg)."
    else:
        reason = f"Es la presentación activa más cercana a {qty_human} kg (objetivo {peso_human} kg)."
    return {
        "source_recipe_id": source_recipe_id,
        "presentacion": best["presentacion"],
        "insumo": best["insumo"],
        "peso": best["peso"],
        "difference": difference,
        "exact_match": exact_match,
        "replacement_quantity": replacement_quantity,
        "reason": reason,
    }


def _recipe_incomplete_erp_item_count(receta: Receta) -> int:
    cached = getattr(receta, "_incomplete_erp_item_count_cache", None)
    if cached is not None:
        return cached
    count = 0
    lineas_qs = receta.lineas.select_related("insumo").only(
        "id",
        "insumo_id",
        "insumo__id",
        "insumo__activo",
        "insumo__tipo_item",
        "insumo__codigo",
        "insumo__categoria",
        "insumo__proveedor_principal_id",
        "insumo__unidad_base_id",
    )
    for linea in lineas_qs:
        if not linea.insumo_id:
            continue
        profile = _insumo_operational_readiness(
            linea.insumo,
            ignore_supplier=receta.tipo == Receta.TIPO_PRODUCTO_FINAL,
        )
        if not profile["ready"]:
            count += 1
    setattr(receta, "_incomplete_erp_item_count_cache", count)
    return count


def _recipe_master_gap_summary(receta: Receta) -> dict[str, object]:
    cached = getattr(receta, "_master_gap_summary_cache", None)
    if cached is not None:
        return cached

    counters = {
        "unidad": 0,
        "proveedor": 0,
        "categoria": 0,
        "codigo_point": 0,
        "inactivo": 0,
    }
    lineas_qs = receta.lineas.select_related("insumo").only(
        "id",
        "insumo_id",
        "insumo__id",
        "insumo__activo",
        "insumo__tipo_item",
        "insumo__codigo",
        "insumo__codigo_point",
        "insumo__categoria",
        "insumo__proveedor_principal_id",
        "insumo__unidad_base_id",
    )
    for linea in lineas_qs:
        if not linea.insumo_id:
            continue
        profile = _insumo_operational_readiness(
            linea.insumo,
            ignore_supplier=receta.tipo == Receta.TIPO_PRODUCTO_FINAL,
        )
        missing = set(profile["missing"])
        if "unidad base" in missing:
            counters["unidad"] += 1
        if "proveedor principal" in missing:
            counters["proveedor"] += 1
        if "categoría" in missing:
            counters["categoria"] += 1
        if "inactivo" in missing:
            counters["inactivo"] += 1
        if linea.insumo.activo and not (linea.insumo.codigo_point or "").strip():
            counters["codigo_point"] += 1

    labels = {
        "unidad": "unidad base",
        "proveedor": "proveedor principal",
        "categoria": "categoría",
        "codigo_point": "código comercial",
        "inactivo": "artículo inactivo",
    }
    dominant_key = None
    dominant_count = 0
    for key, count in counters.items():
        if count > dominant_count:
            dominant_key = key
            dominant_count = count

    summary = {
        "counts": counters,
        "dominant_key": dominant_key,
        "dominant_count": dominant_count,
        "dominant_label": labels.get(dominant_key, ""),
    }
    setattr(receta, "_master_gap_summary_cache", summary)
    return summary


def _linea_counts_as_operational_pending(linea: LineaReceta) -> bool:
    if linea.match_status not in {LineaReceta.STATUS_NEEDS_REVIEW, LineaReceta.STATUS_REJECTED}:
        return False
    if linea.tipo_linea == LineaReceta.TIPO_SUBSECCION and linea.insumo_id:
        return False
    if linea.insumo_id:
        return True
    label = normalizar_nombre(linea.insumo_texto or "")
    return label not in NONBLOCKING_PRODUCT_PLACEHOLDER_LABELS


def _recipe_operational_pending_count(receta: Receta) -> int:
    prefetched = getattr(receta, "_prefetched_objects_cache", {}).get("lineas")
    if prefetched is not None:
        return sum(1 for linea in prefetched if _linea_counts_as_operational_pending(linea))
    lineas_qs = receta.lineas.only("id", "tipo_linea", "match_status", "insumo_id", "insumo_texto")
    return sum(1 for linea in lineas_qs if _linea_counts_as_operational_pending(linea))


def _recipe_has_effective_bom(receta: Receta) -> bool:
    lineas_attr = getattr(receta, "lineas_count", None)
    if lineas_attr is not None and int(lineas_attr or 0) > 0:
        return True
    if lineas_attr is None and receta.lineas.exists():
        return True

    equivalence_attr = getattr(receta, "has_equivalence_bom", None)
    if equivalence_attr is not None:
        if bool(equivalence_attr):
            return True
    elif RecetaEquivalencia.objects.filter(receta_porcion=receta, activo=True).exclude(receta_padre=receta).exists():
        return True

    derived_attr = getattr(receta, "has_derived_bom", None)
    if derived_attr is not None:
        return bool(derived_attr)
    return RecetaPresentacionDerivada.objects.filter(receta_derivada=receta, activo=True).exists()


def _recipe_effective_cost_display(receta: Receta) -> Decimal:
    costo = receta.costo_total_estimado_decimal
    if costo and costo > 0:
        return costo

    equivalence = _recipe_active_equivalence(receta)
    if equivalence is not None:
        parent_cost = equivalence.receta_padre.costo_total_estimado_decimal
        if parent_cost and parent_cost > 0:
            factor = Decimal(str(equivalence.factor_conversion or 1))
            if factor > 0:
                return parent_cost / factor

    relation = _recipe_active_derived_relation(receta)
    if relation is not None:
        parent_cost = relation.receta_padre.costo_total_estimado_decimal
        if parent_cost and parent_cost > 0:
            units = Decimal(str(relation.unidades_por_padre or 0))
            if units > 0:
                return parent_cost / units
    return Decimal("0")


def _recipe_active_equivalence(receta: Receta) -> RecetaEquivalencia | None:
    if hasattr(receta, "_effective_equivalence_cache"):
        return getattr(receta, "_effective_equivalence_cache")
    equivalence = (
        RecetaEquivalencia.objects.select_related("receta_padre")
        .filter(receta_porcion=receta, activo=True)
        .exclude(receta_padre=receta)
        .first()
    )
    setattr(receta, "_effective_equivalence_cache", equivalence)
    return equivalence


def _recipe_active_derived_relation(receta: Receta) -> RecetaPresentacionDerivada | None:
    if hasattr(receta, "_effective_derived_cache"):
        return getattr(receta, "_effective_derived_cache")
    filters = Q(receta_derivada=receta)
    if receta.codigo_point:
        filters |= Q(codigo_point_derivado=receta.codigo_point)
    relation = (
        RecetaPresentacionDerivada.objects.select_related("receta_padre")
        .filter(filters, activo=True)
        .first()
    )
    setattr(receta, "_effective_derived_cache", relation)
    return relation


def _recipe_source_display(receta: Receta) -> str:
    lineas_attr = getattr(receta, "lineas_count", None)
    if (lineas_attr is not None and int(lineas_attr or 0) > 0) or (lineas_attr is None and receta.lineas.exists()):
        return "BOM directo"

    equivalence = _recipe_active_equivalence(receta)
    if equivalence is not None:
        return f"Equivalencia: {equivalence.receta_padre.nombre}"

    relation = _recipe_active_derived_relation(receta)
    if relation is not None:
        return f"Derivada: {relation.receta_padre.nombre}"

    return "Sin costeo"


def _recent_point_sale_window() -> tuple[date | None, date | None]:
    latest_sale = PointDailySale.objects.aggregate(max_date=Max("sale_date")).get("max_date")
    if not latest_sale:
        return None, None
    return latest_sale, latest_sale - timedelta(days=RECENT_PRODUCT_ACTIVITY_DAYS)


def _recent_point_product_activity_snapshot() -> dict[str, object]:
    latest_sale, sale_cutoff = _recent_point_sale_window()
    if not sale_cutoff:
        return {
            "latest_sale": latest_sale,
            "sale_cutoff": sale_cutoff,
            "code_norms": set(),
            "name_norms": set(),
            "receta_ids": set(),
        }

    sales_qs = PointDailySale.objects.filter(sale_date__gte=sale_cutoff)
    code_norms = {
        normalizar_codigo_point(code)
        for code in sales_qs.exclude(product__sku="").order_by().values_list("product__sku", flat=True).distinct()
        if normalizar_codigo_point(code)
    }
    name_norms = {
        name
        for name in sales_qs.exclude(product__normalized_name="")
        .order_by()
        .values_list("product__normalized_name", flat=True)
        .distinct()
        if name
    }
    receta_ids = {
        receta_id
        for receta_id in sales_qs.exclude(receta_id__isnull=True).order_by().values_list("receta_id", flat=True).distinct()
        if receta_id
    }
    return {
        "latest_sale": latest_sale,
        "sale_cutoff": sale_cutoff,
        "code_norms": code_norms,
        "name_norms": name_norms,
        "receta_ids": receta_ids,
    }


def _recipe_has_recent_point_sale(receta: Receta, activity_snapshot: dict[str, object]) -> bool:
    receta_ids = activity_snapshot.get("receta_ids") or set()
    if receta.id in receta_ids:
        return True
    code_norm = normalizar_codigo_point(receta.codigo_point or "")
    if code_norm and code_norm in (activity_snapshot.get("code_norms") or set()):
        return True
    name_norm = normalizar_nombre(receta.nombre or "")
    return bool(name_norm and name_norm in (activity_snapshot.get("name_norms") or set()))


def _excluded_point_category_codes() -> set[str]:
    raw_codes = {
        str(code or "").strip()
        for code in PointProductCategory.objects.filter(
            category__in={
                "SERVICIO_ACCESORIO",
                "REVENTA",
                "TOPPING",
            }
        ).values_list("codigo_point", flat=True)
        if str(code or "").strip()
    }
    excluded_codes = {normalizar_codigo_point(code) for code in raw_codes if normalizar_codigo_point(code)}
    if raw_codes:
        for sku, external_id in PointProduct.objects.filter(
            Q(sku__in=raw_codes) | Q(external_id__in=raw_codes)
        ).values_list("sku", "external_id"):
            for code in (sku, external_id):
                code_norm = normalizar_codigo_point(code or "")
                if code_norm:
                    excluded_codes.add(code_norm)
    return excluded_codes


def _recipe_is_excluded_point_category(receta: Receta, excluded_codes: set[str]) -> bool:
    code_norm = normalizar_codigo_point(receta.codigo_point or "")
    return bool(code_norm and code_norm in excluded_codes)


def _recipe_counts_as_bom_pending(receta: Receta, excluded_point_category_codes: set[str]) -> bool:
    if receta.tipo != Receta.TIPO_PRODUCTO_FINAL:
        return False
    if _recipe_is_excluded_point_category(receta, excluded_point_category_codes):
        return False
    return not _recipe_has_effective_bom(receta)


def _recipe_has_active_point_product(receta: Receta, active_point_codes: set[str]) -> bool:
    return bool(receta.codigo_point and normalizar_codigo_point(receta.codigo_point) in active_point_codes)


def _recipe_master_blockers(
    lineas: list[LineaReceta],
    *,
    receta: Receta | None = None,
    lookback_days: int = 45,
) -> list[dict[str, object]]:
    historico_units = Decimal("0")
    demand_days_count = 0
    if receta and receta.tipo == Receta.TIPO_PRODUCTO_FINAL:
        end_date = timezone.localdate() - timedelta(days=1)
        start_date = end_date - timedelta(days=max(lookback_days - 1, 0))
        historico_qs = VentaHistorica.objects.filter(
            receta_id=receta.id,
            fecha__gte=start_date,
            fecha__lte=end_date,
        )
        historico_units = Decimal(str(historico_qs.aggregate(total=Sum("cantidad"))["total"] or 0))
        demand_days_count = int(historico_qs.values("fecha").distinct().count())

    if historico_units >= Decimal("80") or demand_days_count >= 12:
        priority_label = "Demanda crítica bloqueada"
        priority_tone = "danger"
        priority_detail = "Esta receta ya tiene una señal comercial fuerte y no conviene operar con artículos maestros incompletos."
        is_demand_critical = True
    elif historico_units >= Decimal("30") or demand_days_count >= 5:
        priority_label = "Alta demanda en revisión"
        priority_tone = "warning"
        priority_detail = "La receta ya tiene tracción comercial y conviene cerrar primero estas brechas del maestro."
        is_demand_critical = False
    else:
        priority_label = "Brecha maestra por cerrar"
        priority_tone = "warning"
        priority_detail = "La brecha sigue abierta, aunque todavía sin presión comercial crítica en la ventana observada."
        is_demand_critical = False

    blockers: dict[int, dict[str, object]] = {}
    for linea in lineas:
        if not linea.insumo_id:
            continue
        profile = getattr(linea, "erp_profile", None) or _insumo_erp_readiness(linea.insumo)
        if profile["ready"]:
            continue
        row = blockers.setdefault(
            linea.insumo_id,
            {
                "insumo": linea.insumo,
                "missing": list(profile["missing"]),
                "line_count": 0,
                "line_positions": [],
                "edit_url": reverse("maestros:insumo_update", args=[linea.insumo_id]),
                "historico_units": historico_units.quantize(Decimal("0.1")),
                "demand_days_count": demand_days_count,
                "priority_label": priority_label,
                "priority_tone": priority_tone,
                "priority_detail": priority_detail,
                "is_demand_critical": is_demand_critical,
            },
        )
        row["line_count"] += 1
        row["line_positions"].append(linea.posicion)
    return sorted(
        blockers.values(),
        key=lambda item: (
            0 if item.get("is_demand_critical") else 1,
            -item["line_count"],
            len(item["missing"]),
            (item["insumo"].nombre or "").lower(),
        ),
    )


def _recipe_master_gap_totals_from_blockers(blockers: list[dict[str, object]]) -> dict[str, int]:
    totals = {
        "total": len(blockers),
        "lineas": 0,
        "critical": 0,
        "high_demand": 0,
        "unidad": 0,
        "proveedor": 0,
        "categoria": 0,
        "codigo_point": 0,
        "inactivo": 0,
    }
    for item in blockers:
        missing = set(item.get("missing") or [])
        line_count = int(item.get("line_count") or 0)
        totals["lineas"] += line_count
        if item.get("is_demand_critical"):
            totals["critical"] += 1
        if Decimal(str(item.get("historico_units") or 0)) > 0:
            totals["high_demand"] += 1
        if "unidad base" in missing:
            totals["unidad"] += line_count
        if "proveedor principal" in missing:
            totals["proveedor"] += line_count
        if "categoría" in missing:
            totals["categoria"] += line_count
        if "código comercial" in missing or "código externo" in missing:
            totals["codigo_point"] += line_count
        if "inactivo" in missing:
            totals["inactivo"] += line_count
    return totals


def _recipe_operational_health(receta: Receta) -> dict[str, str]:
    lineas_attr = getattr(receta, "lineas_count", None)
    pendientes = _recipe_operational_pending_count(receta)
    if lineas_attr is None:
        lineas = receta.lineas.count()
    else:
        lineas = int(lineas_attr or 0)
    if receta.tipo == Receta.TIPO_PRODUCTO_FINAL:
        if pendientes > 0:
            return {
                "code": "warning",
                "label": "Por validar",
                "description": "Tiene componentes por validar o artículos estándar todavía sin cerrar.",
            }
        if not _recipe_has_effective_bom(receta):
            return {
                "code": "danger",
                "label": "Incompleta",
                "description": "Aún no tiene componentes de armado.",
            }
        lineas_qs = list(
            receta.lineas.select_related("insumo").only(
                "id",
                "insumo_id",
                "insumo__id",
                "insumo__tipo_item",
                "insumo__codigo",
            )
        )
        upstream_snapshot = getattr(receta, "_product_upstream_snapshot_cache", None)
        if upstream_snapshot is None:
            upstream_snapshot = _product_upstream_snapshot(lineas_qs, receta=receta)
            setattr(receta, "_product_upstream_snapshot_cache", upstream_snapshot)
        direct_base_snapshot = _direct_base_usage_snapshot(lineas_qs)
        if direct_base_snapshot["count"] > 0:
            return {
                "code": "warning",
                "label": "Usa base sin presentación",
                "description": "Está consumiendo la base completa aunque la receta origen ya tiene presentaciones activas.",
            }
        if upstream_snapshot["internal_without_source_count"] > 0:
            return {
                "code": "warning",
                "label": "Sin trazabilidad base",
                "description": "Tiene insumos internos ligados que todavía no apuntan a una receta base canónica.",
            }
        if _recipe_packaging_missing(receta, upstream_snapshot):
            return {
                "code": "warning",
                "label": "Sin empaque",
                "description": "Producto final todavía sin empaque ligado en su BOM.",
            }
        incomplete_items = _recipe_incomplete_erp_item_count(receta)
        if incomplete_items > 0:
            return {
                "code": "warning",
                "label": "Maestro incompleto",
                "description": f"Tiene {incomplete_items} artículo(s) ligados que todavía no están listos en el maestro.",
            }
        return {
            "code": "success",
            "label": "Lista para operar",
            "description": "Producto final con estructura lista para costeo, planeación y operación.",
        }

    if not receta.rendimiento_cantidad or not receta.rendimiento_unidad_id:
        return {
            "code": "danger",
            "label": "Sin rendimiento",
            "description": "Falta rendimiento para costeo enterprise.",
        }
    if receta.usa_presentaciones:
        derived_state = _recipe_derived_sync_state(receta)
        presentaciones_activas = int(derived_state["active_presentaciones"])
        if presentaciones_activas <= 0:
            return {
                "code": "warning",
                "label": "Sin derivados",
                "description": "Está marcada con presentaciones pero no tiene derivados activos.",
            }
        if (not derived_state["prep_ready"]) or int(derived_state["derived_presentaciones"]) < presentaciones_activas:
            return {
                "code": "warning",
                "label": "Sincronizar derivados",
                "description": "Tiene presentaciones activas, pero aún faltan derivados maestros o sincronización de base.",
            }
    incomplete_items = _recipe_incomplete_erp_item_count(receta)
    if incomplete_items > 0:
        return {
            "code": "warning",
            "label": "Maestro incompleto",
            "description": f"Tiene {incomplete_items} artículo(s) ligados que todavía no están listos en el maestro.",
        }
    return {
        "code": "success",
        "label": "Lista para operar",
        "description": "Base preparada para costeo, derivados y uso operativo.",
    }


def _build_recipe_point_validation_snapshot(recetas: list[Receta]) -> dict[int, dict[str, object]]:
    producto_final_ids = [receta.id for receta in recetas if receta.tipo == Receta.TIPO_PRODUCTO_FINAL]
    if not producto_final_ids:
        return {}

    latest_sale = PointDailySale.objects.aggregate(max_date=Max("sale_date")).get("max_date")
    latest_production = PointProductionLine.objects.aggregate(max_date=Max("production_date")).get("max_date")
    sale_cutoff = latest_sale - timedelta(days=90) if latest_sale else None
    production_cutoff = latest_production - timedelta(days=90) if latest_production else None

    sale_code_norms = (
        {
            normalizar_codigo_point(code)
            for code in PointDailySale.objects.filter(sale_date__gte=sale_cutoff)
            .exclude(product__sku="")
            .order_by()
            .values_list("product__sku", flat=True)
            .distinct()
            if normalizar_codigo_point(code)
        }
        if sale_cutoff
        else set()
    )
    sale_name_norms = (
        {
            name
            for name in PointDailySale.objects.filter(sale_date__gte=sale_cutoff)
            .exclude(product__normalized_name="")
            .order_by()
            .values_list("product__normalized_name", flat=True)
            .distinct()
            if name
        }
        if sale_cutoff
        else set()
    )
    production_code_norms = (
        {
            normalizar_codigo_point(code)
            for code in PointProductionLine.objects.filter(production_date__gte=production_cutoff)
            .exclude(item_code="")
            .order_by()
            .values_list("item_code", flat=True)
            .distinct()
            if normalizar_codigo_point(code)
        }
        if production_cutoff
        else set()
    )
    production_name_norms = (
        {
            normalizar_nombre(name)
            for name in PointProductionLine.objects.filter(production_date__gte=production_cutoff)
            .exclude(item_name="")
            .order_by()
            .values_list("item_name", flat=True)
            .distinct()
            if normalizar_nombre(name)
        }
        if production_cutoff
        else set()
    )

    snapshot: dict[int, dict[str, object]] = {}
    for receta in recetas:
        if receta.tipo != Receta.TIPO_PRODUCTO_FINAL:
            continue
        code_norm = normalizar_codigo_point(receta.codigo_point or "")
        name_norm = normalizar_nombre(receta.nombre or "")
        is_candidate = bool(
            (code_norm and code_norm in sale_code_norms)
            or (code_norm and code_norm in production_code_norms)
            or (name_norm and name_norm in sale_name_norms)
            or (name_norm and name_norm in production_name_norms)
        )
        snapshot[receta.id] = {
            "is_candidate": is_candidate,
            "label": "Vigente operativo" if is_candidate else "Archivado operativo",
            "description": (
                "El producto tiene venta o producción reciente en Point."
                if is_candidate
                else "No tiene venta ni producción reciente en Point; no entra al conteo operativo."
            ),
        }
    return snapshot


def _recipe_counts_in_operational_catalog(receta: Receta) -> bool:
    if receta.tipo != Receta.TIPO_PRODUCTO_FINAL:
        return True
    status = getattr(receta, "point_validation_status", None)
    if not status:
        return True
    return bool(status.get("is_candidate", True))


def _module_enablement_handoff_rows(
    receta: Receta,
    module_enablement_cards: list[dict[str, object]],
) -> list[dict[str, object]]:
    fallback_url = reverse("recetas:receta_detail", args=[receta.id])
    meta_map = {
        "Costeo": {
            "owner": "Recetas / Costeo",
            "depends_on": "Maestro + BOM estable",
            "exit_criteria": "Costo trazable y snapshot operativo cerrados.",
        },
        "Costeo base": {
            "owner": "Recetas / Costeo",
            "depends_on": "Rendimiento + unidad",
            "exit_criteria": "La base ya costea con rendimiento y unidad operativa válidos.",
        },
        "Derivados": {
            "owner": "Producción / Recetas",
            "depends_on": "Presentaciones activas",
            "exit_criteria": "Los derivados activos ya existen como artículos internos usables.",
        },
        "Derivados opcionales": {
            "owner": "Producción",
            "depends_on": "Definición operativa",
            "exit_criteria": "La base quedó cerrada para operar sin tamaños derivados.",
        },
        "Producto final": {
            "owner": "Producción / Comercial",
            "depends_on": "Cadena base → derivado",
            "exit_criteria": "La base ya quedó conectada a por lo menos un producto final.",
        },
        "MRP": {
            "owner": "Planeación / Producción",
            "depends_on": "Cadena + costo + maestro",
            "exit_criteria": "La receta ya puede explotar demanda y reabasto sin bloqueos de documento.",
        },
        "MRP / Compras": {
            "owner": "Planeación / Compras",
            "depends_on": "Derivados + consumo final",
            "exit_criteria": "La base ya puede bajar a planeación y abastecimiento aguas abajo.",
        },
        "Compras": {
            "owner": "Compras",
            "depends_on": "Empaque + BOM final",
            "exit_criteria": "Solicitudes y órdenes ya pueden usar este documento sin ambigüedad.",
        },
        "Inventario": {
            "owner": "Inventario / Almacén",
            "depends_on": "Maestro + empaque",
            "exit_criteria": "Inventario ya puede operar el artículo final como referencia estable.",
        },
        "Venta final": {
            "owner": "Operación comercial",
            "depends_on": "Empaque + cadena + costo",
            "exit_criteria": "El artículo ya está listo para operación final consistente.",
        },
        "Uso operativo": {
            "owner": "Producción",
            "depends_on": "Costo + maestro",
            "exit_criteria": "La base ya puede consumirse como insumo interno estable.",
        },
    }
    rows: list[dict[str, object]] = []
    for card in module_enablement_cards:
        label = str(card.get("label") or "Módulo")
        meta = meta_map.get(
            label,
            {
                "owner": "Operación ERP",
                "depends_on": "Cierre documental",
                "exit_criteria": "El documento ya cumple la salida operativa de esta etapa.",
            },
        )
        is_ready = bool(card.get("status"))
        tone = str(card.get("tone") or ("success" if is_ready else "warning"))
        blockers = 0 if is_ready else 1
        rows.append(
            {
                "label": label,
                "owner": meta["owner"],
                "status": "Listo" if is_ready else "Bloqueado",
                "tone": tone,
                "blockers": blockers,
                "completion": 100 if is_ready else 55,
                "depends_on": meta["depends_on"],
                "exit_criteria": meta["exit_criteria"],
                "detail": card.get("detail") or meta["exit_criteria"],
                "next_step": card.get("action_label") or "Abrir documento",
                "url": card.get("action_url") or fallback_url,
                "cta": card.get("action_label") or "Abrir documento",
            }
        )
    return rows


def _matches_recipe_health_filter(receta: Receta, health_filter: str) -> bool:
    health = _recipe_operational_health(receta)
    if health_filter == "listas":
        return health["code"] == "success"
    if health_filter == "pendientes":
        return health["code"] == "warning"
    if health_filter == "incompletas":
        return health["code"] == "danger"
    return True


def _recipe_governance_issues(receta: Receta) -> list[str]:
    issues: list[str] = []
    if not (receta.familia or "").strip():
        issues.append("familia")
    if not (receta.categoria or "").strip():
        issues.append("categoria")
    if _recipe_incomplete_erp_item_count(receta) > 0:
        issues.append("maestro_incompleto")
    if receta.tipo == Receta.TIPO_PREPARACION and (not receta.rendimiento_cantidad or not receta.rendimiento_unidad_id):
        issues.append("rendimiento")
    if receta.tipo == Receta.TIPO_PREPARACION and receta.usa_presentaciones:
        derived_state = _recipe_derived_sync_state(receta)
        presentaciones_activas = int(derived_state["active_presentaciones"])
        if presentaciones_activas <= 0:
            issues.append("derivados")
        elif (not derived_state["prep_ready"]) or int(derived_state["derived_presentaciones"]) < presentaciones_activas:
            issues.append("sync_derivados")
        elif not _recipe_supply_chain_snapshot(receta)["has_downstream_usage"]:
            issues.append("sin_consumo_final")
    if receta.tipo == Receta.TIPO_PRODUCTO_FINAL:
        lineas_count = int(getattr(receta, "lineas_count", 0) or 0)
        if not _recipe_has_effective_bom(receta):
            issues.append("componentes")
        elif lineas_count > 0:
            lineas_qs = list(
                receta.lineas.select_related("insumo").only(
                    "id",
                    "insumo_id",
                    "insumo__id",
                    "insumo__tipo_item",
                    "insumo__codigo",
                )
            )
            upstream_snapshot = getattr(receta, "_product_upstream_snapshot_cache", None)
            if upstream_snapshot is None:
                upstream_snapshot = _product_upstream_snapshot(lineas_qs, receta=receta)
                setattr(receta, "_product_upstream_snapshot_cache", upstream_snapshot)
            direct_base_snapshot = getattr(receta, "direct_base_snapshot", None) or _recipe_direct_base_snapshot(receta)
            receta.direct_base_snapshot = direct_base_snapshot
            if int(direct_base_snapshot["count"]) > 0:
                issues.append("base_directa")
            if _recipe_packaging_missing(receta, upstream_snapshot):
                issues.append("sin_empaque")
            if upstream_snapshot["internal_without_source_count"] > 0:
                issues.append("sin_base_origen")
    return issues


def _recipe_primary_action(receta: Receta) -> dict[str, str] | None:
    issues = getattr(receta, "governance_issues", None) or _recipe_governance_issues(receta)
    if "componentes" in issues:
        return {
            "label": "Cargar estructura",
            "url": reverse("recetas:linea_create", args=[receta.id]),
        }
    if "sin_empaque" in issues:
        return {
            "label": "Agregar empaque",
            "url": f"{reverse('recetas:linea_create', args=[receta.id])}?component_kind=EMPAQUE&component_context=empaques",
        }
    if "sin_base_origen" in issues:
        return {
                "label": "Resolver catálogo",
            "url": f"{reverse('recetas:matching_pendientes')}?receta={receta.id}",
        }
    if "base_directa" in issues:
        direct_base_snapshot = getattr(receta, "direct_base_snapshot", None) or _recipe_direct_base_snapshot(receta)
        if int(direct_base_snapshot["suggested_count"] or 0) > 0:
            return {
                "label": "Aplicar derivados sugeridos",
                "url": reverse("recetas:receta_apply_direct_base_replacements", args=[receta.id]),
            }
        return {
            "label": "Corregir base",
            "url": reverse("recetas:receta_detail", args=[receta.id]),
        }
    if "maestro_incompleto" in issues:
        gap_summary = _recipe_master_gap_summary(receta)
        missing_field = gap_summary["dominant_key"] or "unidad"
        return {
            "label": "Revisar maestro",
            "url": (
                f"{reverse('maestros:insumo_list')}?usage_scope=recipes&recipe_scope=finales"
                f"&enterprise_status=incompletos&missing_field={missing_field}&linked_recipe_id={receta.id}"
            ),
        }
    if "derivados" in issues:
        return {
            "label": "Agregar presentación",
            "url": reverse("recetas:presentacion_create", args=[receta.id]),
        }
    if "sync_derivados" in issues:
        return {
            "label": "Sincronizar derivados",
            "url": f"{reverse('recetas:receta_sync_derivados', args=[receta.id])}?next={reverse('recetas:recetas_list')}",
        }
    if "rendimiento" in issues:
        return {
            "label": "Abrir receta",
            "url": reverse("recetas:receta_detail", args=[receta.id]),
        }
    if "sin_consumo_final" in issues:
        return {
            "label": "Crear producto final",
            "url": f"{reverse('recetas:receta_create')}?mode=FINAL&source_base={receta.id}",
        }
    health = getattr(receta, "operational_health", None) or _recipe_operational_health(receta)
    if health["code"] != "success":
        return {
            "label": "Revisar",
            "url": reverse("recetas:receta_detail", args=[receta.id]),
        }
    return None


def _recipe_chain_checkpoints(receta: Receta) -> list[dict[str, str]]:
    checkpoints: list[dict[str, str]] = []

    if receta.tipo == Receta.TIPO_PREPARACION:
        derived_state = _recipe_derived_sync_state(receta)
        supply = getattr(receta, "supply_chain_snapshot", None) or _recipe_supply_chain_snapshot(receta)

        prep_ready = bool(
            receta.rendimiento_cantidad
            and receta.rendimiento_unidad_id
            and derived_state["prep_ready"]
        )
        checkpoints.append(
            {
                "label": "Base",
                "code": "success" if prep_ready else "danger",
                "detail": "Base maestra lista" if prep_ready else "Falta rendimiento o base maestra",
            }
        )

        if receta.usa_presentaciones:
            deriv_ok = (
                int(derived_state["active_presentaciones"]) > 0
                and int(derived_state["derived_presentaciones"]) >= int(derived_state["active_presentaciones"])
            )
            checkpoints.append(
                {
                    "label": "Derivados",
                    "code": "success" if deriv_ok else "warning",
                    "detail": (
                        f"{derived_state['derived_presentaciones']}/{derived_state['active_presentaciones']} activos"
                    ),
                }
            )

        consumo_ok = bool(supply and supply["has_downstream_usage"])
        checkpoints.append(
            {
                "label": "Uso final",
                "code": "success" if consumo_ok else "warning",
                "detail": (
                    f"{supply['downstream_recipe_count']} producto(s) final(es)"
                    if consumo_ok
                    else "Aún sin consumo final"
                ),
            }
        )
    else:
        upstream_snapshot = getattr(receta, "product_upstream_snapshot", None) or getattr(
            receta, "_product_upstream_snapshot_cache", None
        )
        if upstream_snapshot is None:
            lineas_qs = list(
                receta.lineas.select_related("insumo").only(
                    "id",
                    "insumo_id",
                    "insumo__id",
                    "insumo__tipo_item",
                    "insumo__codigo",
                )
            )
            upstream_snapshot = _product_upstream_snapshot(lineas_qs, receta=receta)
            setattr(receta, "_product_upstream_snapshot_cache", upstream_snapshot)

        internal_ok = int(upstream_snapshot["internal_count"]) > 0
        checkpoints.append(
            {
                "label": "Internos",
                "code": "success" if internal_ok else "warning",
                "detail": (
                    f"{upstream_snapshot['internal_count']} ligado(s)"
                    if internal_ok
                    else "Sin internos"
                ),
            }
        )

        trace_ok = (
            int(upstream_snapshot["internal_without_source_count"]) <= 0
            and int(upstream_snapshot["upstream_base_count"]) > 0
        )
        checkpoints.append(
            {
                "label": "Trazabilidad",
                "code": "success" if trace_ok else "warning",
                "detail": (
                    f"{upstream_snapshot['upstream_base_count']} base(s) origen"
                    if trace_ok
                    else "Sin base origen"
                ),
            }
        )

        pack_ok = _recipe_packaging_ready(receta, upstream_snapshot)
        checkpoints.append(
            {
                "label": "Empaque",
                "code": "success" if pack_ok else "warning",
                "detail": (
                    (
                        f"{upstream_snapshot['empaque_count']} ligado(s)"
                        if int(upstream_snapshot["empaque_count"]) > 0
                        else "Empaque flexible"
                    )
                    if pack_ok
                    else "Sin empaque"
                ),
            }
        )

    for item in checkpoints:
        item["tone"] = {
            "success": "success",
            "warning": "warning",
            "danger": "danger",
        }.get(item["code"], "secondary")
    return checkpoints


def _recipe_failing_chain_checkpoint_codes(receta: Receta) -> set[str]:
    failing: set[str] = set()
    for item in _recipe_chain_checkpoints(receta):
        if item["code"] == "success":
            continue
        if receta.tipo == Receta.TIPO_PREPARACION:
            if item["label"] == "Base":
                failing.add("base_ready")
            elif item["label"] == "Derivados":
                failing.add("derived_sync")
            elif item["label"] == "Uso final":
                failing.add("final_usage")
        else:
            if item["label"] == "Trazabilidad":
                failing.add("upstream_trace")
            elif item["label"] == "Empaque":
                failing.add("packaging_ready")
            elif item["label"] == "Internos":
                failing.add("internal_components")
    return failing


def _recipe_chain_status(receta: Receta) -> dict[str, str]:
    issues = getattr(receta, "governance_issues", None) or _recipe_governance_issues(receta)
    if receta.tipo == Receta.TIPO_PREPARACION:
        if "rendimiento" in issues:
            return {
                "code": "danger",
                "label": "Base incompleta",
                "description": "Falta rendimiento base para costeo y cadena operativa.",
            }
        if "derivados" in issues:
            return {
                "code": "warning",
                "label": "Sin derivados",
                "description": "La base está marcada para presentaciones, pero aún no tiene derivados activos.",
            }
        if "sync_derivados" in issues:
            return {
                "code": "warning",
                "label": "Sincronizar derivados",
                "description": "Derivados en revisión antes del consumo final.",
            }
        if "sin_consumo_final" in issues:
            return {
                "code": "warning",
                "label": "Sin consumo final",
                "description": "La base tiene derivados, pero todavía no alimenta un producto final.",
            }
        return {
            "code": "success",
            "label": "Cadena lista",
            "description": "La base ya puede alimentar derivados y productos finales.",
        }
    if "componentes" in issues:
        return {
            "code": "danger",
            "label": "Sin estructura",
            "description": "El producto final aún no tiene componentes cargados.",
        }
    if "sin_base_origen" in issues:
        return {
            "code": "warning",
            "label": "Sin trazabilidad de base",
            "description": "Hay internos ligados sin trazabilidad a receta base.",
        }
    if "sin_empaque" in issues:
        return {
            "code": "warning",
            "label": "Sin empaque ligado",
            "description": "Producto final todavía sin empaque ligado en su BOM.",
        }
    return {
        "code": "success",
        "label": "Encadenado",
        "description": "Producto final conectado a base interna y empaque.",
    }


def _recipe_chain_actions_catalog(receta: Receta) -> list[dict[str, str]]:
    actions: list[dict[str, str]] = []
    if receta.tipo == Receta.TIPO_PREPARACION:
        derived_state = getattr(receta, "derived_state", None) or _recipe_derived_sync_state(receta)
        supply = getattr(receta, "supply_chain_snapshot", None) or _recipe_supply_chain_snapshot(receta)
        active_presentaciones = int(derived_state["active_presentaciones"])
        derived_presentaciones = int(derived_state["derived_presentaciones"])
        if receta.usa_presentaciones:
            actions.append(
                {"label": "Presentaciones", "url": reverse("recetas:presentacion_create", args=[receta.id])}
            )
            if active_presentaciones <= 0 or (not derived_state["prep_ready"]) or derived_presentaciones < active_presentaciones:
                actions.append(
                    {
                        "label": "Sincronizar derivados",
                        "url": f"{reverse('recetas:receta_sync_derivados', args=[receta.id])}?next={reverse('recetas:recetas_list')}",
                    }
                )
            if supply and not supply["has_downstream_usage"] and derived_presentaciones > 0:
                actions.append(
                    {
                        "label": "Crear final",
                        "url": f"{reverse('recetas:receta_create')}?mode=FINAL&source_base={receta.id}",
                    }
                )
        elif not receta.rendimiento_cantidad:
            actions.append(
                {"label": "Capturar rendimiento", "url": reverse("recetas:receta_detail", args=[receta.id])}
            )
    else:
        upstream_snapshot = getattr(receta, "product_upstream_snapshot", None) or getattr(
            receta, "_product_upstream_snapshot_cache", None
        )
        if upstream_snapshot is None:
            lineas_qs = list(
                receta.lineas.select_related("insumo").only(
                    "id",
                    "insumo_id",
                    "insumo__id",
                    "insumo__tipo_item",
                    "insumo__codigo",
                )
            )
            upstream_snapshot = _product_upstream_snapshot(lineas_qs, receta=receta)
            setattr(receta, "_product_upstream_snapshot_cache", upstream_snapshot)
        if upstream_snapshot["internal_without_source_count"] > 0:
            actions.append(
                {"label": "Resolver referencias", "url": f"{reverse('recetas:matching_pendientes')}?receta={receta.id}"}
            )
        if _recipe_packaging_missing(receta, upstream_snapshot):
            actions.append(
                {
                    "label": "Agregar empaque",
                    "url": f"{reverse('recetas:linea_create', args=[receta.id])}?component_kind=EMPAQUE&component_context=empaques",
                }
            )
    return actions[:3]


def _recipe_chain_focus_summary(receta: Receta) -> dict[str, Any]:
    chain = getattr(receta, "chain_status_info", None) or _recipe_chain_status(receta)
    issues = getattr(receta, "governance_issues", None) or _recipe_governance_issues(receta)
    primary_action = getattr(receta, "primary_action", None) or _recipe_primary_action(receta)
    direct_base_snapshot = getattr(receta, "direct_base_snapshot", None) or _recipe_direct_base_snapshot(receta)
    supply = getattr(receta, "supply_chain_snapshot", None) or _recipe_supply_chain_snapshot(receta)
    upstream_snapshot = getattr(receta, "product_upstream_snapshot", None) or getattr(
        receta, "_product_upstream_snapshot_cache", None
    )
    if upstream_snapshot is None and receta.tipo == Receta.TIPO_PRODUCTO_FINAL:
        lineas_qs = list(
            receta.lineas.select_related("insumo").only(
                "id",
                "insumo_id",
                "insumo__id",
                "insumo__tipo_item",
                "insumo__codigo",
            )
        )
        upstream_snapshot = _product_upstream_snapshot(lineas_qs, receta=receta)
        setattr(receta, "_product_upstream_snapshot_cache", upstream_snapshot)

    if receta.tipo == Receta.TIPO_PREPARACION:
        if "rendimiento" in issues:
            return {
                "tone": "danger",
                "label": "Capturar rendimiento",
                "summary": "La base todavía no puede costearse ni alimentar derivados porque falta rendimiento total o unidad.",
                "action": primary_action,
            }
        if "derivados" in issues:
            return {
                "tone": "warning",
                "label": "Crear presentaciones",
                "summary": "La base está marcada para derivados, pero aún no tiene presentaciones activas para operar.",
                "action": primary_action,
            }
        if "sync_derivados" in issues:
            derived_state = getattr(receta, "derived_state", None) or _recipe_derived_sync_state(receta)
            return {
                "tone": "warning",
                "label": "Sincronizar derivados",
                "summary": (
                    f"La base tiene {derived_state['active_presentaciones']} presentaciones activas y "
                    f"{derived_state['derived_presentaciones']} derivado(s) operativo(s)."
                ),
                "action": primary_action,
            }
        if "sin_consumo_final" in issues:
            return {
                "tone": "warning",
                "label": "Conectar a producto final",
                "summary": (
                    f"La base ya tiene {supply['presentacion_count']} derivado(s) operativos, "
                    "pero todavía no alimenta ningún producto final."
                ),
                "action": primary_action,
            }
        return {
            "tone": "success",
            "label": "Cadena lista",
            "summary": chain["description"],
            "action": primary_action,
        }

    if "componentes" in issues:
        return {
            "tone": "danger",
            "label": "Completar estructura",
            "summary": "El producto final aún no tiene componentes internos cargados para costeo y producción.",
            "action": primary_action,
        }
    if "base_directa" in issues:
        suggested_count = int(direct_base_snapshot.get("suggested_count") or 0)
        return {
            "tone": "warning",
            "label": "Resolver derivados",
            "summary": (
                f"El producto final consume base directa. Hay {suggested_count} reemplazo(s) sugerido(s) listo(s) para aplicar."
                if suggested_count
                else "El producto final consume base directa y aún requiere corrección manual hacia un derivado operativo."
            ),
            "action": primary_action,
        }
    if "sin_base_origen" in issues:
        missing_count = int(upstream_snapshot["internal_without_source_count"] or 0) if upstream_snapshot else 0
        return {
            "tone": "warning",
            "label": "Resolver referencias",
            "summary": f"Hay {missing_count} interno(s) sin trazabilidad a receta base dentro del producto final.",
            "action": primary_action,
        }
    if "sin_empaque" in issues:
        return {
            "tone": "warning",
            "label": "Agregar empaque",
            "summary": "El producto final aún no tiene empaque ligado, por lo que la cadena comercial sigue incompleta.",
            "action": primary_action,
        }
    return {
        "tone": "success",
        "label": "Cadena lista",
        "summary": chain["description"],
        "action": primary_action,
    }


def _recipe_catalog_governance_rows(
    *,
    total_pendientes: int,
    health_summary: dict[str, int],
    master_gap_totals: dict[str, int],
    chain_focus: dict[str, Any],
) -> list[dict[str, object]]:
    master_total = int(master_gap_totals.get("total") or 0)
    incompletas = int(health_summary.get("incompletas") or 0)
    action = chain_focus.get("action") or {}
    return [
        {
            "front": "Catálogo BOM",
            "owner": "Producción / Recetas",
            "blockers": total_pendientes,
            "completion": 100 if total_pendientes == 0 else 70,
            "detail": (
                "El catálogo ya puede operar sin documentos abiertos."
                if total_pendientes == 0
                else f"{total_pendientes} receta(s) siguen en revisión dentro del catálogo."
            ),
            "next_step": (
                "Mantener monitoreo preventivo."
                if total_pendientes == 0
                else "Cerrar brechas estructurales y documentales."
            ),
            "url": reverse("recetas:recetas_list"),
            "cta": "Abrir catálogo",
        },
        {
            "front": "Maestro de artículos",
            "owner": "Maestros / Inventario",
            "blockers": master_total,
            "completion": 100 if master_total == 0 else 75,
            "detail": (
                "El maestro ya no bloquea recetas activas."
                if master_total == 0
                else f"{master_total} artículo(s) ligados siguen incompletos para operar."
            ),
            "next_step": (
                "Mantener integridad maestra."
                if master_total == 0
                else "Completar el maestro antes de seguir liberando documentos."
            ),
            "url": reverse("maestros:insumo_list"),
            "cta": "Abrir maestro",
        },
        {
            "front": "Cadena operativa",
            "owner": "ERP / Operaciones",
            "blockers": incompletas,
            "completion": 100 if incompletas == 0 else 85,
            "detail": chain_focus.get("summary") or "La cadena base → derivado → producto final ya está bajo control.",
            "next_step": (
                chain_focus.get("action_detail")
                or ("Mantener la cadena cerrada." if incompletas == 0 else "Resolver el bloqueo dominante de cadena.")
            ),
            "url": action.get("url") or reverse("recetas:recetas_list"),
            "cta": action.get("label") or "Abrir siguiente paso",
        },
    ]


def _recipe_detail_governance_rows(
    *,
    receta: Receta,
    recipe_master_gap_totals: dict[str, int],
    total_revision: int,
    total_sin_match: int,
    release_gate_progress: dict[str, int],
    chain_focus_summary: dict[str, Any],
) -> list[dict[str, object]]:
    action = chain_focus_summary.get("action") or {}
    master_total = int(recipe_master_gap_totals.get("total") or 0)
    return [
        {
            "front": "Documento BOM",
            "owner": "Producción / Recetas",
            "blockers": total_revision + total_sin_match,
            "completion": int(release_gate_progress.get("pct") or 0),
            "detail": (
                "El documento ya está listo para costeo y operación."
                if (total_revision + total_sin_match) == 0
                else f"{total_revision + total_sin_match} componente(s) siguen abiertos en revisión documental."
            ),
            "next_step": (
                "Mantener monitoreo del documento."
                if (total_revision + total_sin_match) == 0
                else "Cerrar componentes pendientes y referencias abiertas."
            ),
            "url": reverse("recetas:receta_detail", args=[receta.id]),
            "cta": "Abrir documento",
        },
        {
            "front": "Maestro de artículos",
            "owner": "Maestros / Inventario",
            "blockers": master_total,
            "completion": 100 if master_total == 0 else 75,
            "detail": (
                "Los artículos ligados ya están listos para operar."
                if master_total == 0
                else f"{master_total} artículo(s) ligados siguen incompletos en el maestro."
            ),
            "next_step": (
                "Mantener integridad maestra."
                if master_total == 0
                else "Completar datos maestros de los artículos bloqueados."
            ),
            "url": reverse("maestros:insumo_list"),
            "cta": "Abrir maestro",
        },
        {
            "front": "Cadena ERP",
            "owner": "ERP / Operaciones",
            "blockers": 0 if int(release_gate_progress.get("pct") or 0) >= 100 else 1,
            "completion": int(release_gate_progress.get("pct") or 0),
            "detail": chain_focus_summary.get("summary") or "La cadena operativa del documento está bajo control.",
            "next_step": chain_focus_summary.get("action_detail") or "Completar la siguiente acción recomendada.",
            "url": action.get("url") or reverse("recetas:receta_detail", args=[receta.id]),
            "cta": action.get("label") or "Abrir siguiente paso",
        },
    ]


def _recipe_detail_trunk_handoff_rows(
    *,
    receta: Receta,
    recipe_master_gap_totals: dict[str, int],
    total_revision: int,
    total_sin_match: int,
    release_gate_progress: dict[str, int],
    chain_focus_summary: dict[str, Any],
    is_producto_final: bool,
    is_base_con_presentaciones: bool,
    governance_issues: list[str],
    supply_chain_snapshot: dict[str, object] | None,
    product_upstream_snapshot: dict[str, object] | None,
) -> list[dict[str, object]]:
    master_total = int(recipe_master_gap_totals.get("total") or 0)
    document_blockers = total_revision + total_sin_match
    chain_action = chain_focus_summary.get("action") or {}
    chain_blockers = 0 if int(release_gate_progress.get("pct") or 0) >= 100 else 1
    compras_blockers = master_total + chain_blockers
    inventory_blockers = master_total

    if is_producto_final and "sin_empaque" in governance_issues:
        compras_blockers += 1
        inventory_blockers += 1
    if is_base_con_presentaciones and supply_chain_snapshot and not bool(supply_chain_snapshot.get("has_downstream_usage")):
        compras_blockers += 1
    if is_producto_final and product_upstream_snapshot:
        if int(product_upstream_snapshot.get("upstream_base_count") or 0) <= 0:
            compras_blockers += 1

    recetas_ready = (document_blockers + master_total + chain_blockers) == 0
    compras_ready = compras_blockers == 0
    inventario_ready = inventory_blockers == 0

    return [
        {
            "label": "Recetas / BOM",
            "owner": "Producción / Recetas",
            "status": "Listo para operar" if recetas_ready else "Bloqueado",
            "tone": "success" if recetas_ready else "warning",
            "blockers": document_blockers + master_total + chain_blockers,
            "completion": 100 if recetas_ready else max(0, min(95, int(release_gate_progress.get("pct") or 0))),
            "depends_on": "Documento limpio + catálogo ERP + cadena cerrada",
            "exit_criteria": "La receta debe costear, normalizar y operar sin brechas documentales ni artículos abiertos.",
            "detail": (
                "La receta ya puede operar como documento BOM estable."
                if recetas_ready
                else "Todavía hay brechas en documento, catálogo ERP o cadena operativa."
            ),
            "next_step": chain_focus_summary.get("action_detail") or "Cerrar brechas del documento",
            "url": chain_action.get("url") or reverse("recetas:receta_detail", args=[receta.id]),
            "cta": chain_action.get("label") or "Abrir documento",
        },
        {
            "label": "Compras documentales",
            "owner": "Compras / Planeación",
            "status": "Listo para operar" if compras_ready else "Bloqueado",
            "tone": "success" if compras_ready else "warning",
            "blockers": compras_blockers,
            "completion": 100 if compras_ready else max(0, 100 - (compras_blockers * 12)),
            "depends_on": "Producto operable + empaques cerrados + maestro consistente",
            "exit_criteria": "Solicitudes, órdenes y recepciones deben tomar esta receta sin ambigüedad estructural.",
            "detail": (
                "Compras ya puede trabajar este documento como referencia estable."
                if compras_ready
                else "Compras sigue condicionado por maestro, empaque o cadena operativa."
            ),
            "next_step": "Abrir compras" if compras_ready else "Cerrar empaque y catálogo ERP",
            "url": reverse("compras:solicitudes") if compras_ready else reverse("maestros:insumo_list"),
            "cta": "Abrir compras" if compras_ready else "Abrir catálogo ERP",
        },
        {
            "label": "Inventario / Reabasto",
            "owner": "Inventario / Almacén",
            "status": "Listo para operar" if inventario_ready else "Bloqueado",
            "tone": "success" if inventario_ready else "warning",
            "blockers": inventory_blockers,
            "completion": 100 if inventario_ready else max(0, 100 - (inventory_blockers * 15)),
            "depends_on": "Artículos listos + empaque operativo + salida documental estable",
            "exit_criteria": "Inventario debe sostener existencia, stock y reabasto sin artículos abiertos del maestro.",
            "detail": (
                "Inventario ya puede operar esta receta como referencia estable."
                if inventario_ready
                else "Inventario todavía requiere cierre del catálogo ERP o empaque para operar estable."
            ),
            "next_step": "Abrir inventario" if inventario_ready else "Cerrar catálogo ERP",
            "url": reverse("inventario:existencias") if inventario_ready else reverse("maestros:insumo_list"),
            "cta": "Abrir inventario" if inventario_ready else "Abrir catálogo ERP",
        },
    ]


def _recipe_enterprise_stage(receta: Receta) -> dict[str, str]:
    issues = getattr(receta, "governance_issues", None) or _recipe_governance_issues(receta)
    chain = getattr(receta, "chain_status_info", None) or _recipe_chain_status(receta)
    supply = getattr(receta, "supply_chain_snapshot", None) or _recipe_supply_chain_snapshot(receta)
    upstream_snapshot = getattr(receta, "product_upstream_snapshot", None) or getattr(
        receta, "_product_upstream_snapshot_cache", None
    )

    if receta.tipo == Receta.TIPO_PREPARACION:
        if "rendimiento" in issues:
            return {
                "code": "base_setup",
                "label": "Base en configuración",
                "tone": "danger",
                "summary": "Falta rendimiento o unidad para cerrar la batida base.",
            }
        if receta.usa_presentaciones:
            if "derivados" in issues or "sync_derivados" in issues:
                return {
                    "code": "derivados_setup",
                    "label": "Derivados en configuración",
                    "tone": "warning",
                    "summary": "La base ya existe, pero sus presentaciones o derivados aún no están listos.",
                }
            if "sin_consumo_final" in issues:
                return {
                    "code": "ready_for_final",
                    "label": "Base liberada a final",
                    "tone": "warning",
                    "summary": "La base y sus derivados ya están listos, pero todavía no alimentan un producto final.",
                }
            return {
                "code": "feeding_final",
                "label": "Alimentando producto final",
                "tone": "success",
                "summary": "La base ya opera con derivados activos y consumo final.",
            }
        return {
            "code": "base_operativa",
            "label": "Base operativa",
            "tone": "success" if chain["code"] == "success" else "warning",
            "summary": "La receta base ya puede costearse y operar como insumo interno.",
        }

    if "componentes" in issues:
        return {
            "code": "final_setup",
            "label": "Armado inicial",
            "tone": "danger",
            "summary": "El producto final aún no tiene estructura operativa suficiente.",
        }
    if "base_directa" in issues or "sin_base_origen" in issues:
        return {
            "code": "final_normalization",
            "label": "Ajuste de estructura",
            "tone": "warning",
            "summary": "El producto final requiere ajustar bases, derivados u origen de componentes.",
        }
    if "sin_empaque" in issues:
        return {
            "code": "final_packaging",
            "label": "Cierre comercial",
            "tone": "warning",
            "summary": "La estructura productiva está casi completa, pero falta empaque final para cerrar salida comercial.",
        }
    if upstream_snapshot and int(upstream_snapshot.get("upstream_base_count") or 0) > 0:
        return {
            "code": "final_ready",
            "label": "Listo para operación ERP",
            "tone": "success",
            "summary": "El producto final ya está conectado a su cadena base y listo para costeo/abasto.",
        }
    return {
        "code": "final_defined",
        "label": "Producto definido",
        "tone": "success" if chain["code"] == "success" else "warning",
        "summary": "El producto final está estructurado y sin bloqueos críticos visibles.",
    }


def _matches_recipe_enterprise_stage(receta: Receta, selected_stage: str) -> bool:
    return _recipe_enterprise_stage(receta)["code"] == selected_stage


def _recipe_enterprise_stage_playbook(receta: Receta) -> list[dict[str, Any]]:
    stage = _recipe_enterprise_stage(receta)
    checkpoints = _recipe_chain_checkpoints(receta)
    checkpoint_map = {item["label"]: item for item in checkpoints}
    items: list[dict[str, Any]] = []

    if receta.tipo == Receta.TIPO_PREPARACION:
        derived_state = _recipe_derived_sync_state(receta)
        supply = getattr(receta, "supply_chain_snapshot", None) or _recipe_supply_chain_snapshot(receta)
        base_ok = bool(receta.rendimiento_cantidad and receta.rendimiento_unidad_id and derived_state["prep_ready"])
        items.append(
            {
                "label": "Base y rendimiento",
                "done": base_ok,
                "detail": checkpoint_map.get("Base", {}).get("detail", "Sin validar"),
                "action_label": None if base_ok else "Abrir receta",
                "action_url": None if base_ok else reverse("recetas:receta_detail", args=[receta.id]),
            }
        )
        if receta.usa_presentaciones:
            derivados_ok = (
                int(derived_state["active_presentaciones"]) > 0
                and int(derived_state["derived_presentaciones"]) >= int(derived_state["active_presentaciones"])
            )
            items.append(
                {
                    "label": "Presentaciones y derivados",
                    "done": derivados_ok,
                    "detail": checkpoint_map.get("Derivados", {}).get("detail", "Sin validar"),
                    "action_label": None if derivados_ok else "Sincronizar derivados",
                    "action_url": None if derivados_ok else reverse("recetas:receta_sync_derivados", args=[receta.id]),
                }
            )
        uso_final_ok = bool(supply and supply["has_downstream_usage"])
        items.append(
            {
                "label": "Consumo en producto final",
                "done": uso_final_ok,
                "detail": checkpoint_map.get("Uso final", {}).get("detail", "Sin validar"),
                "action_label": None if uso_final_ok else "Crear producto final",
                "action_url": None if uso_final_ok else f"{reverse('recetas:receta_create')}?mode=FINAL&source_base={receta.id}",
            }
        )
    else:
        upstream_snapshot = getattr(receta, "product_upstream_snapshot", None) or getattr(
            receta,
            "_product_upstream_snapshot_cache",
            None,
        )
        if upstream_snapshot is None:
            lineas_qs = list(
                receta.lineas.select_related("insumo").only(
                    "id",
                    "insumo_id",
                    "insumo__id",
                    "insumo__tipo_item",
                    "insumo__codigo",
                )
            )
            upstream_snapshot = _product_upstream_snapshot(lineas_qs, receta=receta)
            setattr(receta, "_product_upstream_snapshot_cache", upstream_snapshot)
        internos_ok = int(upstream_snapshot["internal_count"]) > 0
        items.append(
            {
                "label": "Insumos internos",
                "done": internos_ok,
                "detail": checkpoint_map.get("Internos", {}).get("detail", "Sin validar"),
                "action_label": None if internos_ok else "Agregar interno",
                "action_url": None if internos_ok else f"{reverse('recetas:linea_create', args=[receta.id])}?component_kind=INSUMO_INTERNO&component_context=internos",
            }
        )
        trazabilidad_ok = (
            int(upstream_snapshot["internal_without_source_count"]) <= 0
            and int(upstream_snapshot["upstream_base_count"]) > 0
        )
        items.append(
            {
                "label": "Trazabilidad a base",
                "done": trazabilidad_ok,
                "detail": checkpoint_map.get("Trazabilidad", {}).get("detail", "Sin validar"),
                "action_label": None if trazabilidad_ok else "Revisar estructura",
                "action_url": None if trazabilidad_ok else reverse("recetas:receta_detail", args=[receta.id]),
            }
        )
        empaque_ok = _recipe_packaging_ready(receta, upstream_snapshot)
        items.append(
            {
                "label": "Empaque final",
                "done": empaque_ok,
                "detail": checkpoint_map.get("Empaque", {}).get("detail", "Sin validar"),
                "action_label": None if empaque_ok else "Agregar empaque",
                "action_url": None if empaque_ok else f"{reverse('recetas:linea_create', args=[receta.id])}?component_kind=EMPAQUE&component_context=empaques",
            }
        )

    summary_map = {
        "base_setup": "Completa la base mínima para que el costo y el flujo queden listos.",
        "derivados_setup": "Cierra derivados y presentaciones para habilitar consumo final.",
        "ready_for_final": "La base ya puede alimentar productos finales sin bloqueo.",
        "feeding_final": "La base ya está conectada a productos finales operativos.",
        "base_operativa": "La base simple ya puede costearse y operar como insumo interno.",
        "final_setup": "Completa el rompecabezas del producto final antes de costear o planear.",
        "final_normalization": "Valida artículos y reemplaza bases directas para estabilizar la estructura.",
        "final_packaging": "Agrega empaque final para cerrar el producto en operación.",
        "final_ready": "El producto final está completo para costeo, MRP y compras.",
        "final_defined": "El producto final ya está definido y estable en el catálogo.",
    }
    for item in items:
        item["tone"] = "success" if item["done"] else "warning"
    return [{"label": "Resumen de etapa", "done": stage["tone"] == "success", "detail": summary_map.get(stage["code"], stage["summary"]), "action_label": None, "action_url": None}] + items


def _recipe_stage_progress(playbook: list[dict[str, Any]]) -> dict[str, int]:
    actionable = [item for item in (playbook or []) if item.get("label") != "Resumen de etapa"]
    total = len(actionable)
    completed = sum(1 for item in actionable if item.get("done"))
    pct = int(round((completed / total) * 100)) if total else 100
    return {
        "completed": completed,
        "total": total,
        "pct": pct,
    }


def _recipe_document_status(receta: Receta) -> dict[str, str]:
    stage = getattr(receta, "enterprise_stage", None) or _recipe_enterprise_stage(receta)
    health = getattr(receta, "operational_health", None) or _recipe_operational_health(receta)
    chain = getattr(receta, "chain_status_info", None) or _recipe_chain_status(receta)
    progress = getattr(receta, "enterprise_stage_progress", None) or _recipe_stage_progress(
        getattr(receta, "enterprise_stage_playbook", None) or _recipe_enterprise_stage_playbook(receta)
    )
    lineas_count = int(getattr(receta, "lineas_count", 0) or 0)

    if stage["tone"] == "danger" or health["code"] == "danger":
        return {
            "code": "blocked",
            "label": "Bloqueado",
            "tone": "danger",
            "detail": "Todavía no cumple las reglas mínimas para operar en el ERP.",
        }
    if not _recipe_has_effective_bom(receta):
        return {
            "code": "draft",
            "label": "Borrador",
            "tone": "warning",
            "detail": "Aún no tiene estructura suficiente para validación operativa.",
        }
    if progress["pct"] >= 100 and health["code"] == "success" and chain["code"] == "success":
        return {
            "code": "operable",
            "label": "Operable",
            "tone": "success",
            "detail": "Documento listo para costeo, MRP y abastecimiento.",
        }
    return {
        "code": "validation",
        "label": "Por validar",
        "tone": "warning",
        "detail": "El documento ya arrancó, pero aún requiere cierre operativo.",
    }


def _build_recipe_chain_focus(
    *,
    vista: str,
    chain_summary: dict[str, int],
    checkpoint_summary: dict[str, int],
    governance_summary: dict[str, int],
    filters: dict[str, str],
) -> dict[str, str]:
    base_query = {
        key: value
        for key, value in filters.items()
        if value and key not in {"chain_status", "chain_checkpoint", "governance_issue"}
    }

    def _list_url(**extra: str) -> str:
        query = {**base_query, **{k: v for k, v in extra.items() if v}}
        encoded = urlencode(query)
        return f"{reverse('recetas:recetas_list')}?{encoded}" if encoded else reverse("recetas:recetas_list")

    if vista == "productos":
        priorities = [
            (
                "internal_components",
                max(
                    int(checkpoint_summary["internal_components"] or 0),
                    int(governance_summary.get("componentes") or 0),
                ),
                {
                    "label": "Internos faltantes",
                    "tone": "danger",
                    "summary": "Hay productos finales sin componentes internos ligados en la estructura.",
                    "action_label": "Cargar internos",
                    "action_detail": "Completa la estructura base del producto final antes de costear, producir o abastecer.",
                },
            ),
            (
                "upstream_trace",
                max(
                    int(checkpoint_summary["upstream_trace"] or 0),
                    int(governance_summary.get("sin_base_origen") or 0),
                ),
                {
                    "label": "Trazabilidad en revisión",
                    "tone": "warning",
                    "summary": "Hay productos finales con insumos internos todavía sin base origen identificada.",
                    "action_label": "Resolver referencias",
                    "action_detail": "Resuelve la referencia estándar para ligar cada interno a su batida o receta base.",
                },
            ),
            (
                "packaging_ready",
                max(
                    int(checkpoint_summary["packaging_ready"] or 0),
                    int(governance_summary.get("sin_empaque") or 0),
                ),
                {
                    "label": "Empaque faltante",
                    "tone": "warning",
                    "summary": "Hay productos finales sin empaque ligado, por lo que el costo y la salida no están cerrados.",
                    "action_label": "Agregar empaques",
                    "action_detail": "Completa el empaque final de la estructura para cerrar costeo, surtido y venta.",
                },
            ),
        ]
    elif vista in {"subinsumos", "insumos"}:
        priorities = [
            (
                "base_ready",
                max(
                    int(checkpoint_summary["base_ready"] or 0),
                    int(governance_summary.get("rendimiento") or 0),
                ),
                {
                    "label": "Base por cerrar",
                    "tone": "danger",
                    "summary": "Hay batidas o mezclas base sin rendimiento o sin estructura mínima para operar.",
                    "action_label": "Completar bases",
                    "action_detail": "Captura rendimiento, unidad y base operativa antes de derivar o costear.",
                },
            ),
            (
                "derived_sync",
                max(
                    int(checkpoint_summary["derived_sync"] or 0),
                    int(governance_summary.get("sync_derivados") or 0),
                    int(governance_summary.get("derivados") or 0),
                ),
                {
                    "label": "Sincronización de derivados",
                    "tone": "warning",
                    "summary": "Hay recetas con presentaciones activas que aún no tienen derivados operativos actualizados.",
                    "action_label": "Sincronizar derivados",
                    "action_detail": "Genera o actualiza subinsumos derivados para que producción y costeo consuman artículos correctos.",
                },
            ),
            (
                "final_usage",
                max(
                    int(checkpoint_summary["final_usage"] or 0),
                    int(governance_summary.get("sin_consumo_final") or 0),
                ),
                {
                    "label": "Sin consumo final",
                    "tone": "warning",
                    "summary": "Hay bases ya derivadas que todavía no alimentan ningún producto final.",
                    "action_label": "Crear productos finales",
                    "action_detail": "Conecta la base con sus productos de venta para cerrar la cadena operativa.",
                },
            ),
        ]
    else:
        priorities = [
            (
                "base_ready",
                max(
                    int(checkpoint_summary["base_ready"] or 0),
                    int(governance_summary.get("rendimiento") or 0),
                ),
                {
                    "label": "Base por validar",
                    "tone": "danger",
                    "summary": "Hay recetas base incompletas que están frenando la cadena operativa.",
                    "action_label": "Completar bases",
                    "action_detail": "Corrige rendimiento y estructura base antes de derivar, consumir o costear.",
                },
            ),
            (
                "internal_components",
                max(
                    int(checkpoint_summary["internal_components"] or 0),
                    int(governance_summary.get("componentes") or 0),
                ),
                {
                    "label": "Internos faltantes",
                    "tone": "danger",
                    "summary": "Hay productos finales sin insumos internos suficientes para operar como BOM enterprise.",
                    "action_label": "Cargar internos",
                    "action_detail": "Completa los insumos internos para evitar productos finales incompletos.",
                },
            ),
            (
                "derived_sync",
                max(
                    int(checkpoint_summary["derived_sync"] or 0),
                    int(governance_summary.get("sync_derivados") or 0),
                    int(governance_summary.get("derivados") or 0),
                ),
                {
                    "label": "Sincronización de derivados",
                    "tone": "warning",
                    "summary": "Hay bases con presentaciones activas que aún no han actualizado sus derivados operativos.",
                    "action_label": "Sincronizar derivados",
                    "action_detail": "Alinea bases y subinsumos antes de continuar con armado final.",
                },
            ),
            (
                "packaging_ready",
                max(
                    int(checkpoint_summary["packaging_ready"] or 0),
                    int(governance_summary.get("sin_empaque") or 0),
                ),
                {
                    "label": "Empaque faltante",
                    "tone": "warning",
                    "summary": "Hay productos finales sin empaque y no deberían considerarse listos para salida.",
                    "action_label": "Agregar empaques",
                    "action_detail": "Completa el empaque final para cerrar costo unitario y operación de salida.",
                },
            ),
            (
                "upstream_trace",
                checkpoint_summary["upstream_trace"],
                {
                    "label": "Trazabilidad por validar",
                    "tone": "warning",
                    "summary": "Hay productos finales cuya estructura interna no tiene trazabilidad cerrada a una base origen.",
                    "action_label": "Resolver referencias",
                    "action_detail": "Corrige integración y catálogo estándar para estabilizar costeo, MRP y compras.",
                },
            ),
            (
                "final_usage",
                checkpoint_summary["final_usage"],
                {
                    "label": "Sin consumo final",
                    "tone": "warning",
                    "summary": "Hay bases derivadas que todavía no se consumen en un producto final.",
                    "action_label": "Crear productos finales",
                    "action_detail": "Conecta la base a un producto final para cerrar el flujo enterprise.",
                },
            ),
        ]

    for checkpoint_code, count, meta in priorities:
        if int(count or 0) > 0:
            return {
                "checkpoint": checkpoint_code,
                "count": str(count),
                "label": meta["label"],
                "tone": meta["tone"],
                "summary": meta["summary"],
                "action_label": meta["action_label"],
                "action_url": _list_url(chain_checkpoint=checkpoint_code),
                "action_detail": meta["action_detail"],
            }

    return {
        "checkpoint": "",
        "count": str(chain_summary["listas"]),
        "label": "Cadena operativa lista",
        "tone": "success",
        "summary": "Las recetas visibles ya cumplen la cadena enterprise esperada para su etapa operativa.",
        "action_label": "Abrir recetas listas",
        "action_url": _list_url(chain_status="listas"),
        "action_detail": "Revisa recetas listas para costeo, producción o armado final sin bloqueos de cadena.",
    }


def _reabasto_enterprise_board(
    *,
    recetas_producto: list[Receta],
    inventario_map: dict[int, Any],
    fecha_operacion: date,
) -> dict[str, Any]:
    blocker_cards_map = {
        "sin_inventario": {"key": "sin_inventario", "label": "Sin inventario CEDIS", "tone": "danger", "count": 0},
        "sin_empaque": {"key": "sin_empaque", "label": "Sin empaque", "tone": "warning", "count": 0},
        "maestro_incompleto": {"key": "maestro_incompleto", "label": "Maestro incompleto", "tone": "warning", "count": 0},
        "pendiente": {"key": "pendiente", "label": "Receta por validar", "tone": "danger", "count": 0},
        "lista": {"key": "lista", "label": "Lista para operar", "tone": "success", "count": 0},
    }
    detail_rows: list[dict[str, Any]] = []
    current_url = f"{reverse('recetas:reabasto_cedis')}?fecha={fecha_operacion.isoformat()}"
    blocked_recipe_ids: set[int] = set()
    for receta in recetas_producto:
        health = _recipe_operational_health(receta)
        blockers: list[dict[str, str]] = []
        if health["label"] == "Sin empaque":
            blockers.append(
                {
                    "key": "sin_empaque",
                    "label": "Sin empaque",
                    "detail": health["description"],
                    "action_label": "Agregar empaque",
                    "action_url": f"{reverse('recetas:linea_create', args=[receta.id])}?component_kind=EMPAQUE&component_context=empaques",
                }
            )
        elif health["label"] == "Maestro incompleto":
            blockers.append(
                {
                    "key": "maestro_incompleto",
                    "label": "Maestro incompleto",
                    "detail": health["description"],
                    "action_label": "Abrir receta",
                    "action_url": reverse("recetas:receta_detail", args=[receta.id]),
                }
            )
        elif health["code"] != "success":
            blockers.append(
                {
                    "key": "pendiente",
                    "label": "Receta por validar",
                    "detail": health["description"],
                    "action_label": "Abrir receta",
                    "action_url": reverse("recetas:receta_detail", args=[receta.id]),
                }
            )
        if receta.id not in inventario_map:
            blockers.append(
                {
                    "key": "sin_inventario",
                    "label": "Sin inventario CEDIS",
                    "detail": "El producto no tiene stock cargado en inventario CEDIS para abastecimiento.",
                    "action_label": "Registrar inventario",
                    "action_url": f"{current_url}#inventario-cedis",
                }
            )

        if blockers:
            blocked_recipe_ids.add(receta.id)
            seen_blocker_keys: set[str] = set()
            for blocker in blockers:
                if blocker["key"] in seen_blocker_keys:
                    continue
                blocker_cards_map[blocker["key"]]["count"] += 1
                seen_blocker_keys.add(blocker["key"])
            for index, blocker in enumerate(blockers):
                detail_rows.append(
                    {
                        "blocker_key": blocker["key"],
                        "receta_id": receta.id,
                        "receta_nombre": receta.nombre,
                        "blocker_label": blocker["label"],
                        "detail": blocker["detail"],
                        "action_label": blocker["action_label"],
                        "action_url": blocker["action_url"],
                        "is_primary": index == 0,
                    }
                )
        else:
            blocker_cards_map["lista"]["count"] += 1

    blocker_cards = [
        item for item in blocker_cards_map.values() if item["count"] > 0 or item["label"] == "Listas"
    ]
    blocker_cards.sort(key=lambda item: (item["label"] == "Listas", -int(item["count"]), item["label"]))
    detail_rows.sort(key=lambda item: (item["blocker_label"], item["receta_nombre"]))
    return {
        "blocker_cards": blocker_cards,
        "detail_rows": detail_rows[:12],
        "blocked_total": len(blocked_recipe_ids),
        "ready_total": blocker_cards_map["lista"]["count"],
    }


def _find_reabasto_plan(fecha_operacion: date) -> PlanProduccion | None:
    marker = f"[AUTO_REABASTO_CEDIS:{fecha_operacion.isoformat()}]"
    nombre_plan = f"CEDIS Reabasto {fecha_operacion.isoformat()}"
    return (
        PlanProduccion.objects.filter(
            fecha_produccion=fecha_operacion,
        )
        .filter(Q(nombre=nombre_plan) | Q(notas__icontains=marker))
        .order_by("-id")
        .first()
    )


def _reabasto_enterprise_context(fecha_operacion: date) -> dict[str, Any]:
    recetas_producto = list(
        Receta.objects.filter(tipo=Receta.TIPO_PRODUCTO_FINAL)
        .order_by("nombre")
        .only("id", "nombre", "codigo_point")
    )
    if not recetas_producto:
        recetas_producto = list(Receta.objects.order_by("nombre").only("id", "nombre", "codigo_point")[:400])

    inventario_cedis = list(
        InventarioCedisProducto.objects.select_related("receta")
        .filter(receta__in=[r.id for r in recetas_producto])
        .order_by("receta__nombre")
    )
    inventario_map = {inv.receta_id: inv for inv in inventario_cedis}
    productos_sin_inventario = [r for r in recetas_producto if r.id not in inventario_map]
    reabasto_enterprise_board = _reabasto_enterprise_board(
        recetas_producto=recetas_producto,
        inventario_map=inventario_map,
        fecha_operacion=fecha_operacion,
    )
    return {
        "recetas_producto": recetas_producto,
        "inventario_cedis": inventario_cedis,
        "inventario_map": inventario_map,
        "productos_sin_inventario": productos_sin_inventario,
        "reabasto_enterprise_board": reabasto_enterprise_board,
    }


def _reabasto_master_blockers(plan: PlanProduccion | None) -> dict[str, Any]:
    if not plan:
        return {
            "cards": [],
            "detail_rows": [],
            "focus": None,
            "focus_rows": [],
        }

    referencia = f"PLAN_PRODUCCION:{plan.id}"
    solicitudes = list(
        SolicitudCompra.objects.select_related("insumo", "insumo__unidad_base")
        .filter(area=referencia, insumo__isnull=False)
        .order_by("insumo__nombre", "id")
    )
    ordenes = list(
        OrdenCompra.objects.select_related("solicitud__insumo", "solicitud__insumo__unidad_base")
        .filter(referencia=referencia, solicitud__insumo__isnull=False)
        .order_by("solicitud__insumo__nombre", "id")
    )
    recepciones = list(
        RecepcionCompra.objects.select_related("orden__solicitud__insumo", "orden__solicitud__insumo__unidad_base")
        .filter(orden__referencia=referencia, orden__solicitud__insumo__isnull=False)
        .order_by("orden__solicitud__insumo__nombre", "id")
    )

    insumos_by_id: dict[int, Insumo] = {}
    for solicitud in solicitudes:
        if solicitud.insumo_id:
            insumos_by_id[solicitud.insumo_id] = solicitud.insumo
    for orden in ordenes:
        insumo = getattr(getattr(orden, "solicitud", None), "insumo", None)
        if insumo and insumo.id:
            insumos_by_id[insumo.id] = insumo
    for recepcion in recepciones:
        insumo = getattr(getattr(getattr(recepcion, "orden", None), "solicitud", None), "insumo", None)
        if insumo and insumo.id:
            insumos_by_id[insumo.id] = insumo

    groups: dict[str, dict[str, Any]] = {}
    detail_rows: list[dict[str, Any]] = []
    for insumo in sorted(insumos_by_id.values(), key=lambda item: (item.nombre or "").lower()):
        profile = _insumo_erp_readiness(insumo)
        if profile["ready"]:
            continue
        article_class = _insumo_article_class(insumo)
        class_key = article_class["key"]
        class_label = article_class["label"]
        group = groups.setdefault(
            class_key,
            {
                "class_key": class_key,
                "class_label": class_label,
                "count": 0,
                "missing_totals": defaultdict(int),
            },
        )
        group["count"] += 1
        missing_fields = list(profile["missing"])
        for missing_label in missing_fields:
            group["missing_totals"][missing_label] += 1

        primary_missing = missing_fields[0] if missing_fields else None
        action_meta = _enterprise_blocker_action_meta_for_recipes(
            insumo.nombre,
            class_key,
            primary_missing,
            insumo_id=insumo.id,
            usage_scope="recipes",
        )
        detail_rows.append(
            {
                "class_key": class_key,
                "class_label": class_label,
                "insumo_nombre": insumo.nombre,
                "name": insumo.nombre,
                "missing_field": _missing_field_to_filter_key(primary_missing or "") or "maestro",
                "missing": ", ".join(missing_fields),
                "detail": "Completa el maestro para habilitar compras, recepción y abastecimiento CEDIS del plan.",
                "action_label": action_meta["label"],
                "action_detail": action_meta["detail"],
                "action_url": action_meta["url"],
                "edit_url": action_meta["edit_url"],
                "tone": "warning",
            }
        )

    cards: list[dict[str, Any]] = []
    for group in sorted(groups.values(), key=lambda item: (-item["count"], item["class_label"])):
        dominant_label = ""
        dominant_count = 0
        for missing_label, count in dict(group["missing_totals"]).items():
            if count > dominant_count:
                dominant_label = missing_label
                dominant_count = count
        action_label, action_detail = _enterprise_blocker_label_detail_for_missing_recipes(dominant_label)
        query = {
            "tipo_item": group["class_key"],
            "enterprise_status": "incompletos",
            "usage_scope": "recipes",
        }
        filter_key = _missing_field_to_filter_key(dominant_label)
        if filter_key:
            query["missing_field"] = filter_key
        cards.append(
            {
                "key": group["class_key"],
                "class_key": group["class_key"],
                "class_label": group["class_label"],
                "count": group["count"],
                "dominant_label": dominant_label or "maestro incompleto",
                "dominant_count": dominant_count,
                "action_label": action_label,
                "action_detail": action_detail,
                "action_url": reverse("maestros:insumo_list") + f"?{urlencode(query)}",
            }
        )

    focus_rows = list(detail_rows[:3])
    if focus_rows:
        first_focus = focus_rows[0]
        focus = {
            **first_focus,
            "label": f"{first_focus['class_label']} · {first_focus['missing_field']}",
            "summary": (
                f"El flujo CEDIS sigue condicionado por {first_focus['insumo_nombre']} "
                f"({first_focus['missing_field']})."
            ),
            "tone": "warning",
        }
    else:
        focus = {
            "class_label": "Maestro",
            "missing_field": "sin bloqueos",
            "label": "Maestro ERP al día",
            "summary": "No hay artículos del plan bloqueando el flujo CEDIS por faltantes de maestro.",
            "action_label": "Abrir maestro",
            "action_detail": "Puedes revisar el catálogo general para seguimiento preventivo.",
            "action_url": reverse("maestros:insumo_list"),
            "tone": "success",
            "count": 0,
        }

    return {
        "cards": cards[:6],
        "detail_rows": detail_rows[:12],
        "focus": focus,
        "focus_rows": focus_rows,
    }


def _reabasto_daily_control(
    *,
    fecha_operacion: date,
    resumen_cierre: dict[str, Any],
    reabasto_enterprise_board: dict[str, Any],
    consolidado_rows: list[dict[str, Any]],
    stage_key: str = "auto",
    closure_key: str = "auto",
    handoff_key: str = "auto",
    master_focus_key: str = "auto",
) -> dict[str, Any]:
    plan = _find_reabasto_plan(fecha_operacion)
    plan_items = int(plan.items.count()) if plan else 0
    demand_history_summary = _reabasto_demand_history_summary(fecha_operacion)
    master_demand_gate = _plan_master_demand_gate(plan) if plan else None
    master_demand_rows = list((master_demand_gate or {}).get("rows") or [])[:3]
    doc_control = (
        _plan_document_control(
            plan,
            stage_key=stage_key,
            closure_key=closure_key,
            handoff_key=handoff_key,
        )
        if plan
        else None
    )
    generation_gate = _reabasto_generation_gate(
        fecha_operacion=fecha_operacion,
        resumen_cierre=resumen_cierre,
        reabasto_enterprise_board=reabasto_enterprise_board,
        consolidado_rows=consolidado_rows,
        plan=plan,
        doc_control=doc_control,
        demand_history_summary=demand_history_summary,
        master_demand_gate=master_demand_gate,
    )

    stage_label = "Esperando cierres"
    stage_tone = "danger"
    stage_detail = "Todavía faltan cierres o existen sucursales tardías para arrancar producción CEDIS."
    next_action_label = "Revisar cierres"
    next_action_url = "#cierre-sucursales"

    if resumen_cierre["listo_8am"]:
        if (master_demand_gate or {}).get("blockers"):
            stage_label = "Demanda crítica bloqueada"
            stage_tone = "danger"
            stage_detail = str((master_demand_gate or {}).get("detail") or "El reabasto no debe liberarse mientras el maestro crítico siga abierto.")
            next_action_label = str((master_demand_gate or {}).get("action_label") or "Cerrar prioridad crítica")
            next_action_url = str((master_demand_gate or {}).get("action_url") or reverse("maestros:insumo_list"))
        elif reabasto_enterprise_board["blocked_total"] > 0:
            stage_label = "Bloqueos operativos"
            stage_tone = "warning"
            stage_detail = "Hay productos finales bloqueados por receta, empaque, inventario o maestro."
            next_action_label = "Revisar bloqueos"
            next_action_url = "#bloqueos-abastecimiento"
        elif not plan:
            stage_label = "Listo para plan"
            stage_tone = "primary"
            stage_detail = "Los cierres están completos y no hay bloqueos críticos; ya puede generarse el plan CEDIS."
            next_action_label = "Generar plan"
            next_action_url = "#filtro-operativo"
        elif doc_control:
            stage_label = doc_control["stage_label"]
            stage_tone = doc_control["stage_tone"]
            stage_detail = doc_control["stage_detail"]
            next_action_label = doc_control["next_action_label"]
            next_action_url = doc_control["next_action_url"]

    control_cards = [
        {
            "label": "Sucursales pendientes",
            "count": int(resumen_cierre["pendientes"] + resumen_cierre["tardias"]),
            "detail": f"Por validar {resumen_cierre['pendientes']} · Tardías {resumen_cierre['tardias']}",
            "tone": "danger" if (resumen_cierre["pendientes"] + resumen_cierre["tardias"]) else "success",
            "action_label": "Abrir cierres",
            "action_url": "#cierre-sucursales",
        },
        {
            "label": "Productos bloqueados",
            "count": int(reabasto_enterprise_board["blocked_total"]),
            "detail": f"Listos {reabasto_enterprise_board['ready_total']}",
            "tone": "warning" if reabasto_enterprise_board["blocked_total"] else "success",
            "action_label": "Abrir bloqueos",
            "action_url": "#bloqueos-abastecimiento",
        },
        {
            "label": "Plan CEDIS",
            "count": plan_items,
            "detail": plan.nombre if plan else "Sin plan generado",
            "tone": "primary" if plan else "warning",
            "action_label": "Abrir plan" if plan else "Generar plan",
            "action_url": f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}" if plan else "#filtro-operativo",
        },
        {
            "label": "Faltante a producir",
            "count": _quantize_qty(generation_gate["shortage_total"]),
            "detail": f"Solicitado { _quantize_qty(generation_gate['requested_total']) }",
            "tone": "primary" if generation_gate["shortage_total"] > 0 else "success",
            "action_label": "Ver consolidado",
            "action_url": "#consolidado-cedis",
        },
    ]
    if (master_demand_gate or {}).get("blockers"):
        control_cards.insert(
            2,
            {
                "label": "Demanda crítica bloqueada",
                "count": int((master_demand_gate or {}).get("blockers") or 0),
                "detail": str((master_demand_gate or {}).get("detail") or "El maestro crítico sigue abierto."),
                "tone": "danger",
                "action_label": str((master_demand_gate or {}).get("action_label") or "Cerrar prioridad crítica"),
                "action_url": str((master_demand_gate or {}).get("action_url") or reverse("maestros:insumo_list")),
            },
        )

    document_cards: list[dict[str, Any]] = []
    document_stage_rows: list[dict[str, Any]] = []
    document_blocker_rows: list[dict[str, Any]] = []
    pipeline_steps: list[dict[str, Any]] = []
    purchase_gate: dict[str, Any] | None = None
    stage_focus: dict[str, Any] | None = None
    closure_checks: list[dict[str, Any]] = []
    closure_focus: dict[str, Any] | None = None
    closure_focus_rows: list[dict[str, Any]] = []
    handoff_checks: list[dict[str, Any]] = []
    handoff_focus: dict[str, Any] | None = None
    handoff_focus_rows: list[dict[str, Any]] = []
    master_blocker_class_cards: list[dict[str, Any]] = []
    master_blocker_detail_rows: list[dict[str, Any]] = []
    master_focus: dict[str, Any] | None = None
    master_focus_rows: list[dict[str, Any]] = []
    master_summary: dict[str, Any] = {
        "label": "Maestro ERP al día",
        "tone": "success",
        "ready_count": 0,
        "blocked_count": 0,
        "progress_pct": 100,
        "detail": "No hay artículos bloqueando el reabasto CEDIS por faltantes del maestro.",
    }
    trunk_handoff_rows: list[dict[str, Any]] = []
    selected_master_focus_key = "auto"
    reabasto_master_board = _reabasto_master_blockers(plan)
    if doc_control:
        document_cards = doc_control["document_cards"]
        document_stage_rows = doc_control["document_stage_rows"]
        document_blocker_rows = doc_control["document_blocker_rows"]
        pipeline_steps = doc_control["pipeline_steps"]
        purchase_gate = doc_control["purchase_gate"]
        stage_focus = doc_control["stage_focus"]
        closure_checks = doc_control["closure_checks"]
        closure_focus = doc_control["closure_focus"]
        closure_focus_rows = doc_control.get("closure_focus_rows") or []
        handoff_checks = doc_control["handoff_checks"]
        handoff_focus = doc_control["handoff_focus"]
        handoff_focus_rows = doc_control.get("handoff_focus_rows") or []
        master_cards = list(reabasto_master_board.get("cards") or [])
        valid_master_focus_keys = {str(card.get("key") or "").strip() for card in master_cards if card.get("key")}
        selected_master_focus_key = master_focus_key if master_focus_key in valid_master_focus_keys else "auto"
        focus_base = reverse("recetas:reabasto_cedis") + f"?{urlencode({'fecha': fecha_operacion.isoformat()})}"
        for card in master_cards:
            card_key = str(card.get("key") or "").strip()
            card["focus_url"] = f"{focus_base}&master_focus_key={urlencode({'k': card_key})[2:]}"
            card["is_active"] = selected_master_focus_key != "auto" and card_key == selected_master_focus_key
        master_detail_source = list(reabasto_master_board.get("detail_rows") or [])
        master_blocker_detail_rows = (
            [row for row in master_detail_source if str(row.get("class_key") or "").strip() == selected_master_focus_key]
            if selected_master_focus_key != "auto"
            else master_detail_source
        )
        master_blocker_class_cards = master_cards
        master_focus_rows = list(master_blocker_detail_rows[:3])
        master_focus = reabasto_master_board.get("focus")
        if master_focus and selected_master_focus_key != "auto" and master_focus_rows:
            first_master_focus = master_focus_rows[0]
            master_focus = {
                **master_focus,
                **first_master_focus,
                "label": f"{first_master_focus['class_label']} · {first_master_focus['missing_field']}",
                "summary": (
                    f"El flujo CEDIS sigue condicionado por {first_master_focus['insumo_nombre']} "
                    f"({first_master_focus['missing_field']})."
                ),
            }
        master_blocked_total = sum(int(card.get("count") or 0) for card in master_cards) or len(master_detail_source)
        master_ready_total = 0 if master_blocked_total else int(reabasto_master_board.get("ready_for_purchase_total") or 0)
        master_total = master_ready_total + master_blocked_total
        master_progress_pct = int(round((master_ready_total / master_total) * 100)) if master_total else 100
        if master_blocked_total:
            dominant_card = master_cards[0] if master_cards else None
            dominant_text = ""
            if dominant_card:
                dominant_text = (
                    f" Bloqueo dominante: {dominant_card.get('class_label', 'Artículo')} · "
                    f"{dominant_card.get('dominant_label', 'maestro incompleto')}."
                )
            master_summary = {
                "label": "Maestro ERP con bloqueos",
                "tone": "warning",
                "ready_count": master_ready_total,
                "blocked_count": master_blocked_total,
                "progress_pct": master_progress_pct,
                "detail": (
                    f"{master_blocked_total} artículo(s) siguen condicionando el reabasto por datos incompletos."
                    f"{dominant_text}"
                ),
            }
        else:
            master_summary = {
                "label": "Maestro ERP al día",
                "tone": "success",
                "ready_count": master_ready_total,
                "blocked_count": 0,
                "progress_pct": master_progress_pct,
                "detail": "El maestro no está bloqueando el reabasto CEDIS para esta fecha.",
            }

    generation_blockers = sum(1 for item in generation_gate.get("checks", []) if not item.get("is_ready"))
    critical_master_open = bool((master_demand_gate or {}).get("blockers") or 0)
    plan_ready = bool(plan and generation_gate.get("can_generate_plan"))
    compras_ready = bool(plan and generation_gate.get("can_generate_compras"))
    inventario_ready = bool(
        plan
        and int(reabasto_enterprise_board["blocked_total"]) == 0
        and int(master_summary.get("blocked_count") or 0) == 0
    )
    trunk_handoff_rows = [
        {
            "label": "Sucursales / Plan",
            "owner": "Ventas / Planeación / Producción",
            "status": "Crítico" if critical_master_open else "Listo para operar" if plan_ready else "Bloqueado",
            "tone": "danger" if critical_master_open else "success" if plan_ready else "warning",
            "blockers": generation_blockers,
            "completion": 12 if critical_master_open else 100 if plan_ready else max(0, 100 - (generation_blockers * 12)),
            "depends_on": "Cierres completos + faltante real + plan generado",
            "exit_criteria": "CEDIS debe contar con cierres completos y un plan de producción válido para arrancar.",
            "detail": (
                "CEDIS no debe liberar el día mientras siga abierta la demanda crítica bloqueada por maestro."
                if critical_master_open
                else "CEDIS ya tiene cierres completos y plan operativo para arrancar."
                if plan_ready
                else "Todavía faltan cierres, faltante validado o plan generado para arrancar el día."
            ),
            "next_step": str((master_demand_gate or {}).get("next_step") or next_action_label) if critical_master_open else next_action_label,
            "url": str((master_demand_gate or {}).get("action_url") or next_action_url) if critical_master_open else next_action_url,
            "cta": str((master_demand_gate or {}).get("action_label") or next_action_label) if critical_master_open else next_action_label,
        },
        {
            "label": "Compras documentales",
            "owner": "Compras / Almacén",
            "status": "Crítico" if critical_master_open else "Listo para operar" if compras_ready else "Bloqueado",
            "tone": "danger" if critical_master_open else "success" if compras_ready else "warning",
            "blockers": int((purchase_gate or {}).get("blocked") or 0) if purchase_gate else 1,
            "completion": 12 if critical_master_open else 100 if compras_ready else int(master_summary.get("progress_pct") or 0),
            "depends_on": "Solicitudes + órdenes + recepciones sin bloqueo",
            "exit_criteria": "El abastecimiento documental debe quedar emitido y sin bloqueos para surtir a tiempo.",
            "detail": (
                "Compras no debe avanzar mientras siga abierta la demanda crítica bloqueada por maestro."
                if critical_master_open
                else "El flujo documental ya puede operar para abastecer el reabasto del día."
                if compras_ready
                else (purchase_gate or {}).get("detail") or "Compras sigue condicionado por bloqueos documentales o de maestro."
            ),
            "next_step": str((master_demand_gate or {}).get("next_step") or (purchase_gate or {}).get("cta_label") or "Abrir plan") if critical_master_open else (purchase_gate or {}).get("cta_label") or "Abrir plan",
            "url": str((master_demand_gate or {}).get("action_url") or (purchase_gate or {}).get("cta_url") or (f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}" if plan else "#filtro-operativo")) if critical_master_open else (purchase_gate or {}).get("cta_url") or (f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}" if plan else "#filtro-operativo"),
            "cta": str((master_demand_gate or {}).get("action_label") or (purchase_gate or {}).get("cta_label") or "Abrir plan") if critical_master_open else (purchase_gate or {}).get("cta_label") or "Abrir plan",
        },
        {
            "label": "Inventario / Reabasto",
            "owner": "Inventario / CEDIS",
            "status": "Listo para operar" if inventario_ready else "Bloqueado",
            "tone": "success" if inventario_ready else "warning",
            "blockers": int(reabasto_enterprise_board["blocked_total"]) + int(master_summary.get("blocked_count") or 0),
            "completion": 100 if inventario_ready else max(0, 100 - ((int(reabasto_enterprise_board["blocked_total"]) + int(master_summary.get("blocked_count") or 0) or 1) * 10)),
            "depends_on": "Productos desbloqueados + maestro al día + inventario CEDIS cargado",
            "exit_criteria": "CEDIS debe poder producir y surtir sin bloqueos de receta, empaque o maestro.",
            "detail": (
                "Inventario y reabasto ya pueden sostener la operación diaria de CEDIS."
                if inventario_ready
                else "CEDIS todavía tiene productos bloqueados o artículos incompletos en maestro."
            ),
            "next_step": "Ver bloqueos" if not inventario_ready else "Abrir consolidado",
            "url": "#bloqueos-abastecimiento" if not inventario_ready else "#consolidado-cedis",
            "cta": "Ver bloqueos" if not inventario_ready else "Abrir consolidado",
        },
    ]

    return {
        "stage_label": stage_label,
        "stage_tone": stage_tone,
        "stage_detail": stage_detail,
        "next_action_label": next_action_label,
        "next_action_url": next_action_url,
        "plan": plan,
        "control_cards": control_cards,
        "document_cards": document_cards,
        "document_stage_rows": document_stage_rows,
        "document_blocker_rows": document_blocker_rows,
        "pipeline_steps": pipeline_steps,
        "purchase_gate": purchase_gate,
        "stage_focus": stage_focus,
        "closure_checks": closure_checks,
        "closure_focus": closure_focus,
        "closure_focus_rows": closure_focus_rows,
        "handoff_checks": handoff_checks,
        "handoff_focus": handoff_focus,
        "handoff_focus_rows": handoff_focus_rows,
        "master_blocker_class_cards": master_blocker_class_cards,
        "master_blocker_detail_rows": master_blocker_detail_rows,
        "master_focus": master_focus,
        "master_focus_rows": master_focus_rows,
        "master_summary": master_summary,
        "master_demand_gate": master_demand_gate,
        "master_demand_rows": master_demand_rows,
        "selected_master_focus_key": selected_master_focus_key,
        "generation_gate": generation_gate,
        "trunk_handoff_rows": trunk_handoff_rows,
        "trunk_handoff_summary": _trunk_handoff_summary(
            trunk_handoff_rows,
            owner="CEDIS / Compras / Inventario",
            fallback_url=reverse("recetas:reabasto_cedis") + f"?{urlencode({'fecha': fecha_operacion.isoformat()})}",
        ),
    }


def _reabasto_daily_decisions(
    *,
    demand_history_summary: dict[str, Any] | None,
    daily_control: dict[str, Any],
) -> list[dict[str, object]]:
    decisions: list[dict[str, object]] = []
    generation_gate = dict(daily_control.get("generation_gate") or {})
    demand_gate = dict(generation_gate.get("demand_gate") or {})
    master_demand_gate = dict(daily_control.get("master_demand_gate") or {})
    purchase_gate = dict(daily_control.get("purchase_gate") or {})

    def push(priority: int, tone: str, title: str, detail: str, url: str, cta: str) -> None:
        decisions.append(
            {
                "priority": priority,
                "tone": tone,
                "title": title,
                "detail": detail,
                "url": url,
                "cta": cta,
            }
        )

    if master_demand_gate and int(master_demand_gate.get("blockers") or 0) > 0:
        push(
            100,
            str(master_demand_gate.get("tone") or "danger"),
            "Cerrar maestro crítico del reabasto",
            str(master_demand_gate.get("detail") or "Hay artículos críticos bloqueando el plan CEDIS."),
            str(master_demand_gate.get("action_url") or reverse("maestros:insumo_list")),
            str(master_demand_gate.get("action_label") or "Abrir maestro"),
        )

    if demand_gate and str(demand_gate.get("tone") or "") == "danger":
        push(
            95,
            "danger",
            "Validar base comparable de demanda",
            str(demand_gate.get("detail") or "La base comparable no es suficiente para generar el reabasto con confianza."),
            str(demand_gate.get("action_url") or "#historico-demanda"),
            str(demand_gate.get("action_label") or "Revisar demanda"),
        )

    if demand_history_summary and int(demand_history_summary.get("sample_days") or 0) == 0:
        push(
            90,
            "danger",
            "Construir referencia histórica",
            "No hay días comparables previos para soportar el reabasto. Conviene revisar captura y demanda antes de generar plan.",
            "#historico-demanda",
            "Ver histórico",
        )

    if str(daily_control.get("stage_tone") or "") == "danger":
        push(
            85,
            str(daily_control.get("stage_tone") or "danger"),
            "Completar cierres de sucursal",
            str(daily_control.get("stage_detail") or "El arranque CEDIS sigue esperando cierres de sucursal."),
            str(daily_control.get("next_action_url") or reverse("recetas:reabasto_cedis_captura")),
            str(daily_control.get("next_action_label") or "Abrir cierres"),
        )

    if purchase_gate and str(purchase_gate.get("tone") or "") != "success":
        push(
            80,
            str(purchase_gate.get("tone") or "warning"),
            "Destrabar flujo de compras",
            str(purchase_gate.get("detail") or "El flujo documental de compras sigue abierto o bloqueado para este reabasto."),
            str(purchase_gate.get("cta_url") or reverse("compras:solicitudes")),
            str(purchase_gate.get("cta_label") or "Abrir compras"),
        )

    if not decisions:
        push(
            10,
            "success",
            "Reabasto sin alertas dominantes",
            "Cierres, demanda, maestro y compras están en condición operable para el arranque CEDIS.",
            reverse("recetas:reabasto_cedis"),
            "Actualizar reabasto",
        )

    decisions.sort(key=lambda item: int(item.get("priority") or 0), reverse=True)
    return decisions[:5]


def _reabasto_branch_priority_rows(
    *,
    fecha_operacion: date,
    sucursales: list[Sucursal],
    resumen_cierre: dict[str, Any],
    demand_history_summary: dict[str, Any] | None,
) -> list[dict[str, object]]:
    solicitud_map = {
        item.sucursal_id: item
        for item in SolicitudReabastoCedis.objects.filter(fecha_operacion=fecha_operacion, sucursal__in=sucursales)
        .select_related("sucursal")
    }
    demand_branch_map = {
        row.get("sucursal__codigo"): row
        for row in list((demand_history_summary or {}).get("top_branches") or [])
    }
    cierre_map = {row["sucursal"].id: row for row in list(resumen_cierre.get("detalle") or [])}

    rows: list[dict[str, object]] = []
    for sucursal in sucursales:
        cierre = cierre_map.get(sucursal.id) or {}
        solicitud = solicitud_map.get(sucursal.id)
        demand_row = demand_branch_map.get(sucursal.codigo) or {}
        lineas = list(solicitud.lineas.all()) if solicitud else []
        solicitado_total = sum((_to_decimal_safe(linea.solicitado) for linea in lineas), Decimal("0"))
        sugerido_total = sum((_to_decimal_safe(linea.sugerido) for linea in lineas), Decimal("0"))
        dominant_line = None
        dominant_units = Decimal("0")
        for linea in lineas:
            candidate_units = max(_to_decimal_safe(linea.solicitado), _to_decimal_safe(linea.sugerido))
            if candidate_units > dominant_units:
                dominant_line = linea
                dominant_units = candidate_units
        pendiente_envio = bool(cierre.get("semaforo") in {"rojo", "amarillo"})
        demand_units = Decimal(str(demand_row.get("total") or 0))
        if pendiente_envio and demand_units > 0:
            tone = "danger"
            status = "Prioridad crítica"
            detail = f"Cierre pendiente/tardío con {demand_units:.0f} unidades históricas comparables."
            priority_score = demand_units + Decimal("200")
        elif pendiente_envio:
            tone = "warning"
            status = "Cierre pendiente"
            detail = "La sucursal todavía no deja un cierre confiable para el arranque CEDIS."
            priority_score = Decimal("120")
        elif sugerido_total > 0 or solicitado_total > 0:
            tone = "primary"
            status = "Lista para surtir"
            detail = f"Sugerido { _quantize_qty(sugerido_total) } · Solicitado { _quantize_qty(solicitado_total) }."
            priority_score = max(sugerido_total, solicitado_total)
        else:
            tone = "success"
            status = "Sin presión visible"
            detail = "No trae cierre atrasado ni faltante sugerido relevante para hoy."
            priority_score = Decimal("0")

        rows.append(
            {
                "sucursal_codigo": sucursal.codigo,
                "sucursal_nombre": sucursal.nombre,
                "status": status,
                "tone": tone,
                "detail": detail,
                "historico_units": demand_units,
                "sugerido_total": _quantize_qty(sugerido_total),
                "solicitado_total": _quantize_qty(solicitado_total),
                "dominant_recipe_id": int(getattr(dominant_line, "receta_id", 0) or 0),
                "dominant_recipe_name": (
                    getattr(getattr(dominant_line, "receta", None), "nombre", "") or "Producto"
                ),
                "dominant_recipe_units": _quantize_qty(dominant_units),
                "action_url": _reabasto_redirect(fecha_operacion, sucursal.id, capture_only=False),
                "action_label": "Abrir sucursal",
                "priority_score": priority_score,
            }
        )

    tone_order = {"danger": 0, "warning": 1, "primary": 2, "success": 3}
    rows.sort(
        key=lambda item: (
            tone_order.get(str(item.get("tone") or ""), 9),
            -float(item.get("priority_score") or 0),
            str(item.get("sucursal_codigo") or ""),
        )
    )
    return rows[:6]


def _reabasto_branch_supply_rows(
    *,
    fecha_operacion: date,
    branch_priority_rows: list[dict[str, object]],
    limit: int = 6,
) -> list[dict[str, object]]:
    if not branch_priority_rows:
        return []

    recipe_ids = {
        int(row.get("dominant_recipe_id") or 0)
        for row in branch_priority_rows
        if int(row.get("dominant_recipe_id") or 0) > 0
    }
    if not recipe_ids:
        return []

    lineas = list(
        LineaReceta.objects.filter(receta_id__in=recipe_ids, insumo_id__isnull=False)
        .exclude(tipo_linea=LineaReceta.TIPO_SUBSECCION)
        .select_related("insumo", "receta", "insumo__unidad_base")
    )
    if not lineas:
        return []

    canonical_by_line: dict[int, Insumo] = {}
    canonical_ids: set[int] = set()
    lineas_by_recipe: dict[int, list[LineaReceta]] = defaultdict(list)
    for linea in lineas:
        if not linea.insumo:
            continue
        canonical = canonical_insumo(linea.insumo) or linea.insumo
        canonical_by_line[linea.id] = canonical
        canonical_ids.add(canonical.id)
        lineas_by_recipe[int(linea.receta_id)].append(linea)

    existencia_map = {
        int(item.insumo_id): item
        for item in ExistenciaInsumo.objects.filter(insumo_id__in=canonical_ids).select_related("insumo")
    }

    rows: list[dict[str, object]] = []
    for branch_row in branch_priority_rows:
        recipe_id = int(branch_row.get("dominant_recipe_id") or 0)
        recipe_units = _to_decimal_safe(branch_row.get("dominant_recipe_units"))
        if recipe_id <= 0 or recipe_units <= 0:
            continue

        best_candidate: dict[str, object] | None = None
        best_score = Decimal("-1")
        for linea in lineas_by_recipe.get(recipe_id, []):
            canonical = canonical_by_line.get(linea.id)
            if canonical is None:
                continue

            required_qty = _to_decimal_safe(linea.cantidad) * recipe_units
            if required_qty <= 0:
                continue

            existencia = existencia_map.get(canonical.id)
            stock_actual = _to_decimal_safe(getattr(existencia, "stock_actual", 0))
            shortage = max(required_qty - stock_actual, Decimal("0"))
            readiness = _insumo_erp_readiness(canonical)
            missing = list(readiness.get("missing") or [])
            latest_cost = _latest_cost_for_insumo(canonical)
            missing_cost = latest_cost is None
            score = (shortage * Decimal("100")) + (Decimal(str(len(missing))) * Decimal("50")) + required_qty
            if missing_cost:
                score += Decimal("25")
            if score > best_score:
                best_score = score
                best_candidate = {
                    "insumo_nombre": canonical.nombre,
                    "required_qty": _quantize_qty(required_qty),
                    "stock_actual": _quantize_qty(stock_actual),
                    "shortage": _quantize_qty(shortage),
                    "master_missing": missing,
                    "missing_cost": missing_cost,
                    "unidad": canonical.unidad_base.codigo if getattr(canonical, "unidad_base_id", None) else "-",
                    "action_url": reverse("maestros:insumo_update", args=[canonical.id]),
                    "action_label": "Asegurar insumo",
                }

        if not best_candidate:
            continue

        rows.append(
            {
                "sucursal_codigo": branch_row.get("sucursal_codigo") or "",
                "sucursal_nombre": branch_row.get("sucursal_nombre") or "Sucursal",
                "dominant_recipe_name": branch_row.get("dominant_recipe_name") or "Producto",
                "insumo_nombre": best_candidate["insumo_nombre"],
                "required_qty": best_candidate["required_qty"],
                "stock_actual": best_candidate["stock_actual"],
                "shortage": best_candidate["shortage"],
                "master_missing": best_candidate["master_missing"],
                "missing_cost": best_candidate["missing_cost"],
                "unidad": best_candidate["unidad"],
                "action_url": best_candidate["action_url"],
                "action_label": best_candidate["action_label"],
                "priority_score": best_score,
            }
        )

    rows.sort(
        key=lambda item: (
            Decimal(str(item.get("shortage") or 0)),
            Decimal(str(len(item.get("master_missing") or []))),
            Decimal(str(item.get("required_qty") or 0)),
        ),
        reverse=True,
    )
    return rows[:limit]


def _reabasto_generation_gate(
    *,
    fecha_operacion: date,
    resumen_cierre: dict[str, Any],
    reabasto_enterprise_board: dict[str, Any],
    consolidado_rows: list[dict[str, Any]],
    plan: PlanProduccion | None,
    doc_control: dict[str, Any] | None,
    demand_history_summary: dict[str, Any] | None = None,
    master_demand_gate: dict[str, Any] | None = None,
) -> dict[str, Any]:
    requested_total = sum((_to_decimal_safe(row.get("total_solicitado")) for row in consolidado_rows), Decimal("0"))
    shortage_total = sum((_to_decimal_safe(row.get("cedis_faltante_producir")) for row in consolidado_rows), Decimal("0"))
    doc_blocked_total = int((doc_control or {}).get("blocked_total") or 0)
    doc_total = 0
    if doc_control:
        doc_total = int(doc_control.get("solicitudes_total") or 0) + int(doc_control.get("ordenes_total") or 0) + int(doc_control.get("recepciones_total") or 0)
    demand_gate = _commercial_signal_gate(
        demand_history_summary,
        context_label="el reabasto diario",
        action_url="#historico-demanda",
        action_label="Revisar base histórica",
    )
    master_demand_ready = not bool((master_demand_gate or {}).get("blockers") or 0)

    checks = [
        {
            "label": "Cierres de sucursal completos",
            "is_ready": bool(resumen_cierre["listo_8am"]),
            "detail": (
                "Todas las sucursales enviaron en tiempo para arranque 8:00 AM."
                if resumen_cierre["listo_8am"]
                else f"Por atender {resumen_cierre['pendientes']} · Tardías {resumen_cierre['tardias']}"
            ),
            "action_label": "Abrir cierres",
            "action_url": "#cierre-sucursales",
        },
        {
            "label": "Productos finales sin bloqueo",
            "is_ready": int(reabasto_enterprise_board["blocked_total"]) == 0,
            "detail": (
                "No hay bloqueos por receta, empaque, inventario ni maestro."
                if int(reabasto_enterprise_board["blocked_total"]) == 0
                else f"Bloqueadas {reabasto_enterprise_board['blocked_total']} · Listas {reabasto_enterprise_board['ready_total']}"
            ),
            "action_label": "Ver bloqueos",
            "action_url": "#bloqueos-abastecimiento",
        },
        {
            "label": "Base histórica comparable",
            "is_ready": bool(demand_gate["is_ready"]),
            "detail": str(demand_gate["detail"]),
            "action_label": str(demand_gate["action_label"]),
            "action_url": str(demand_gate["action_url"]),
        },
        {
            "label": "Maestro crítico del plan cerrado",
            "is_ready": master_demand_ready,
            "detail": (
                "No hay artículos críticos por demanda bloqueando el plan CEDIS."
                if master_demand_ready
                else str((master_demand_gate or {}).get("detail") or "El plan sigue bloqueado por artículos críticos del maestro.")
            ),
            "action_label": str((master_demand_gate or {}).get("action_label") or "Cerrar prioridad crítica"),
            "action_url": str((master_demand_gate or {}).get("action_url") or reverse("maestros:insumo_list")),
        },
        {
            "label": "Faltante a producir detectado",
            "is_ready": shortage_total > 0,
            "detail": (
                f"Faltante CEDIS { _quantize_qty(shortage_total) }"
                if shortage_total > 0
                else "No hay faltante a producir; no corresponde generar plan."
            ),
            "action_label": "Ver consolidado",
            "action_url": "#consolidado-cedis",
        },
        {
            "label": "Plan CEDIS generado",
            "is_ready": bool(plan),
            "detail": (
                f"{plan.nombre} · {int(plan.items.count())} renglones"
                if plan
                else "Todavía no existe plan para esta fecha."
            ),
            "action_label": "Abrir plan" if plan else "Generar plan",
            "action_url": f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}" if plan else "#filtro-operativo",
        },
        {
            "label": "Flujo de compras sin bloqueos",
            "is_ready": bool(plan) and doc_total > 0 and doc_blocked_total == 0,
            "detail": (
                "Sin documentos bloqueados en solicitudes, órdenes o recepciones."
                if plan and doc_total > 0 and doc_blocked_total == 0
                else f"Bloqueos documentales {doc_blocked_total}"
                if plan and doc_total > 0
                else (
                    "Todavía no existen documentos de compras para este plan."
                    if plan
                    else "Primero debe existir un plan CEDIS antes de generar compras."
                )
            ),
            "action_label": (
                (doc_control or {}).get("next_action_label")
                if doc_control
                else ("Abrir plan" if plan else "Generar plan")
            ),
            "action_url": (
                (doc_control or {}).get("next_action_url")
                if doc_control
                else (f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}" if plan else "#filtro-operativo")
            ),
        },
    ]

    can_generate_plan = bool(checks[0]["is_ready"] and checks[1]["is_ready"] and checks[2]["is_ready"] and checks[3]["is_ready"] and checks[4]["is_ready"])
    can_generate_compras = bool(can_generate_plan and checks[6]["is_ready"])

    if can_generate_compras:
        overall_label = "Listo para generar"
        overall_tone = "success"
        overall_detail = "Puede generarse o regenerarse el flujo documental de compras desde el faltante consolidado."
    elif can_generate_plan:
        overall_label = "Listo para plan"
        overall_tone = "primary"
        overall_detail = "El plan puede generarse, pero compras sigue condicionado por bloqueos documentales existentes."
    else:
        overall_label = "Bloqueado"
        overall_tone = "danger"
        overall_detail = "Primero corrige cierres, bloqueos operativos o faltante inexistente antes de generar documentos."

    return {
        "requested_total": requested_total,
        "shortage_total": shortage_total,
        "doc_total": doc_total,
        "checks": checks,
        "can_generate_plan": can_generate_plan,
        "can_generate_compras": can_generate_compras,
        "overall_label": overall_label,
        "overall_tone": overall_tone,
        "overall_detail": overall_detail,
        "demand_gate": demand_gate,
    }


def _reabasto_generation_blocker_message(generation_gate: dict[str, Any], *, target: str) -> str:
    failing = [item["label"] for item in generation_gate.get("checks", []) if not item.get("is_ready")]
    if target == "compras" and generation_gate.get("can_generate_plan") and not generation_gate.get("can_generate_compras"):
        return "No se puede generar compras todavía: el plan está listo, pero el flujo documental actual tiene bloqueos abiertos."
    if failing:
        return f"No se puede generar {target} todavía: {', '.join(failing)}."
    return f"No se puede generar {target} todavía por validaciones enterprise en revisión."


def _log_reabasto_gate_block(user, *, fecha_operacion: date, target: str, generation_gate: dict[str, Any]) -> None:
    failing = [item["label"] for item in generation_gate.get("checks", []) if not item.get("is_ready")]
    log_event(
        user,
        "BLOCKED",
        "recetas.ReabastoCedis",
        None,
        {
            "source": "REABASTO_CEDIS_GATE",
            "target": target,
            "fecha_operacion": fecha_operacion.isoformat(),
            "overall_label": generation_gate.get("overall_label"),
            "overall_tone": generation_gate.get("overall_tone"),
            "requested_total": float(_to_decimal_safe(generation_gate.get("requested_total"))),
            "shortage_total": float(_to_decimal_safe(generation_gate.get("shortage_total"))),
            "doc_total": int(generation_gate.get("doc_total") or 0),
            "failing_checks": failing,
        },
    )


def _recipes_critical_path_rows(
    rows: list[dict[str, object]],
    *,
    owner: str = "Recetas / Operación",
    fallback_url: str,
) -> list[dict[str, object]]:
    severity_order = {"danger": 0, "warning": 1, "success": 2, "primary": 3}
    normalized_rows: list[dict[str, object]] = []
    for row in rows:
        open_count = int(
            row.get("open_count")
            if row.get("open_count") is not None
            else row.get("open")
            if row.get("open") is not None
            else row.get("blocked")
            if row.get("blocked") is not None
            else row.get("count")
            or 0
        )
        completion = int(
            row.get("progress_pct")
            if row.get("progress_pct") is not None
            else row.get("completion")
            if row.get("completion") is not None
            else 0
        )
        tone = str(row.get("semaphore_tone") or row.get("tone") or ("danger" if open_count else "success"))
        normalized_rows.append(
            {
                "title": row.get("label") or row.get("title") or row.get("front") or "Tramo de recetas",
                "owner": row.get("owner") or owner,
                "status": row.get("status") or row.get("semaphore_label") or ("En revisión" if open_count else "Cerrado"),
                "tone": tone,
                "count": open_count,
                "completion": completion,
                "depends_on": row.get("depends_on") or ("Tramo previo del documento" if row.get("step") else "Inicio del flujo"),
                "dependency_status": row.get("dependency_status") or row.get("detail") or "Sin dependencia registrada",
                "detail": row.get("detail", ""),
                "next_step": row.get("next_step") or row.get("cta") or row.get("action_detail") or "Revisar tramo",
                "url": row.get("action_url") or row.get("action_href") or row.get("url") or fallback_url,
                "cta": row.get("action_label") or row.get("cta") or "Abrir",
            }
        )

    ranked = sorted(
        normalized_rows,
        key=lambda item: (
            severity_order.get(str(item.get("tone") or "warning"), 9),
            -int(item.get("count") or 0),
            int(item.get("completion") or 0),
        ),
    )
    critical_rows: list[dict[str, object]] = []
    for index, item in enumerate(ranked[:4], start=1):
        critical_rows.append(
            {
                "rank": f"R{index}",
                "title": item["title"],
                "owner": item["owner"],
                "status": item["status"],
                "tone": item["tone"],
                "count": item["count"],
                "completion": item["completion"],
                "depends_on": item["depends_on"],
                "dependency_status": item["dependency_status"],
                "detail": item["detail"],
                "next_step": item["next_step"],
                "url": item["url"],
                "cta": item["cta"],
            }
        )
    return critical_rows


def _recipes_executive_radar_rows(
    rows: list[dict[str, object]],
    *,
    owner: str = "Recetas / Operación",
    fallback_url: str,
) -> list[dict[str, object]]:
    radar_rows: list[dict[str, object]] = []
    for row in rows[:4]:
        blockers = int(
            row.get("open_count")
            if row.get("open_count") is not None
            else row.get("open")
            if row.get("open") is not None
            else row.get("blocked")
            if row.get("blocked") is not None
            else row.get("count")
            or 0
        )
        completion = int(
            row.get("progress_pct")
            if row.get("progress_pct") is not None
            else row.get("completion")
            if row.get("completion") is not None
            else 0
        )
        tone = str(row.get("semaphore_tone") or row.get("tone") or ("danger" if blockers else "success"))
        if blockers <= 0 and completion >= 90:
            status = "Controlado"
            dominant_blocker = "Sin bloqueo activo"
        elif completion >= 50:
            status = "En seguimiento"
            dominant_blocker = row.get("detail", "") or "Brecha operativa en seguimiento"
        else:
            status = "Con bloqueo"
            dominant_blocker = row.get("detail", "") or "Bloqueo operativo abierto"
        radar_rows.append(
            {
                "phase": row.get("label") or row.get("title") or row.get("front") or "Frente de recetas",
                "owner": row.get("owner") or owner,
                "status": status,
                "tone": tone,
                "blockers": blockers,
                "progress_pct": completion,
                "dominant_blocker": dominant_blocker,
                "depends_on": row.get("depends_on") or "Inicio del flujo",
                "dependency_status": row.get("dependency_status") or row.get("detail") or "Sin dependencia registrada",
                "next_step": row.get("next_step") or row.get("cta") or row.get("action_detail") or "Revisar tramo",
                "url": row.get("action_url") or row.get("action_href") or row.get("url") or fallback_url,
                "cta": row.get("action_label") or row.get("cta") or "Abrir",
            }
        )
    return radar_rows


def _trunk_handoff_summary(
    rows: list[dict[str, object]],
    *,
    owner: str,
    fallback_url: str,
) -> dict[str, object] | None:
    if not rows:
        return None
    normalized: list[dict[str, object]] = []
    for row in rows:
        blockers = int(
            row.get("blockers")
            if row.get("blockers") is not None
            else row.get("open_count")
            if row.get("open_count") is not None
            else row.get("open")
            if row.get("open") is not None
            else row.get("count")
            or 0
        )
        completion = int(
            row.get("completion")
            if row.get("completion") is not None
            else row.get("progress_pct")
            if row.get("progress_pct") is not None
            else 0
        )
        normalized.append(
            {
                "label": row.get("label") or row.get("title") or "Tramo ERP",
                "blockers": blockers,
                "completion": completion,
                "next_step": row.get("next_step") or row.get("cta") or "Revisar tramo",
                "url": row.get("url") or row.get("action_url") or fallback_url,
                "cta": row.get("cta") or row.get("action_label") or "Abrir",
            }
        )
    avg_completion = int(round(sum(int(item["completion"]) for item in normalized) / len(normalized)))
    total_blockers = sum(int(item["blockers"]) for item in normalized)
    priority = sorted(normalized, key=lambda item: (-int(item["blockers"]), int(item["completion"])))[0]
    tone = "success" if total_blockers == 0 and avg_completion >= 100 else "warning" if total_blockers == 0 else "danger"
    status = "Controlado" if tone == "success" else "Seguimiento" if tone == "warning" else "Crítico"
    return {
        "owner": owner,
        "status": status,
        "tone": tone,
        "completion": avg_completion,
        "blockers": total_blockers,
        "next_step": priority["next_step"],
        "url": priority["url"],
        "cta": priority["cta"],
        "priority_label": priority["label"],
        "segments": len(normalized),
    }


def _plan_daily_decisions(
    *,
    plan_actual: PlanProduccion | None,
    demand_gate_summary: dict[str, Any] | None,
    master_demand_gate_summary: dict[str, Any] | None,
    ventas_historicas_summary: dict[str, Any] | None,
    document_control: dict[str, Any] | None,
) -> list[dict[str, object]]:
    decisions: list[dict[str, object]] = []

    def push(priority: int, tone: str, title: str, detail: str, url: str, cta: str) -> None:
        decisions.append(
            {
                "priority": priority,
                "tone": tone,
                "title": title,
                "detail": detail,
                "url": url,
                "cta": cta,
            }
        )

    if master_demand_gate_summary and int(master_demand_gate_summary.get("blockers") or 0) > 0:
        push(
            100,
            str(master_demand_gate_summary.get("tone") or "danger"),
            "Cerrar maestro crítico del plan",
            str(master_demand_gate_summary.get("detail") or "Hay artículos maestros críticos reteniendo la liberación del plan."),
            str(master_demand_gate_summary.get("action_url") or reverse("maestros:insumo_list")),
            str(master_demand_gate_summary.get("action_label") or "Abrir maestro"),
        )

    if demand_gate_summary and str(demand_gate_summary.get("tone") or "") == "danger":
        push(
            95,
            "danger",
            "Validar base comercial del plan",
            str(demand_gate_summary.get("detail") or "La base comercial del plan es frágil."),
            str(demand_gate_summary.get("action_url") or "#plan-pronosticos"),
            str(demand_gate_summary.get("action_label") or "Abrir pronóstico"),
        )

    if ventas_historicas_summary and int(ventas_historicas_summary.get("missing_days") or 0) > 0:
        push(
            88,
            "warning",
            "Completar histórico de ventas",
            (
                f"Faltan {ventas_historicas_summary.get('missing_days', 0)} días dentro del rango histórico. "
                "Eso debilita el soporte estadístico del plan."
            ),
            "#plan-pronosticos",
            "Ver base estadística",
        )

    if not plan_actual:
        push(
            92,
            "danger",
            "Crear plan operativo",
            "No hay un plan seleccionado. Primero define el documento rector antes de explotar demanda y compras.",
            reverse("recetas:plan_produccion"),
            "Crear plan",
        )

    if document_control and int(document_control.get("blocked_total") or 0) > 0:
        push(
            82,
            "warning",
            "Destrabar flujo documental",
            f"Hay {document_control.get('blocked_total', 0)} bloqueos documentales en solicitudes, órdenes o recepciones del plan.",
            str(document_control.get("next_action_url") or reverse("compras:solicitudes")),
            str(document_control.get("next_action_label") or "Abrir compras"),
        )

    if document_control and int(document_control.get("recepciones_abiertas_total") or 0) > 0:
        push(
            75,
            "warning",
            "Cerrar recepciones abiertas",
            f"Hay {document_control.get('recepciones_abiertas_total', 0)} recepciones abiertas que siguen impactando el cierre operativo del plan.",
            reverse("compras:recepciones"),
            "Abrir recepciones",
        )

    if not decisions:
        push(
            10,
            "success",
            "Plan sin alertas dominantes",
            "El plan tiene base comercial, maestro y flujo documental en condición operable.",
            reverse("recetas:plan_produccion"),
            "Actualizar plan",
        )

    decisions.sort(key=lambda item: int(item.get("priority") or 0), reverse=True)
    return decisions[:5]


def _plan_branch_priority_rows(
    *,
    plan_actual: PlanProduccion | None,
    periodo: str,
) -> list[dict[str, object]]:
    if not plan_actual:
        return []

    recipe_ids = list(plan_actual.items.values_list("receta_id", flat=True))
    if not recipe_ids:
        return []

    month = None
    try:
        month = int((periodo or "").split("-")[1])
    except (IndexError, TypeError, ValueError):
        month = None

    try:
        sales_qs = VentaHistorica.objects.filter(receta_id__in=recipe_ids, sucursal_id__isnull=False)
        if month:
            monthly_qs = sales_qs.filter(fecha__month=month)
            sales_qs = monthly_qs if monthly_qs.exists() else sales_qs
        sales_rows = list(
            sales_qs.values("sucursal_id", "sucursal__codigo", "sucursal__nombre")
            .annotate(
                total_units=Sum("cantidad"),
                total_tickets=Sum("tickets"),
                recipe_count=Count("receta_id", distinct=True),
            )
            .order_by("-total_units", "sucursal__codigo")
        )
        request_rows = list(
            SolicitudVenta.objects.filter(
                periodo=periodo,
                receta_id__in=recipe_ids,
                sucursal_id__isnull=False,
            )
            .values("sucursal_id")
            .annotate(
                solicitud_total=Sum("cantidad"),
                solicitud_rows=Count("id"),
            )
        )
    except (OperationalError, ProgrammingError):
        return []

    request_map = {int(row["sucursal_id"]): row for row in request_rows if row.get("sucursal_id")}
    recipe_plan_map = {
        int(item.receta_id): _to_decimal_safe(item.cantidad)
        for item in plan_actual.items.select_related("receta").all()
    }
    branch_recipe_rows = list(
        sales_qs.values("sucursal_id", "receta_id", "receta__nombre")
        .annotate(total_units=Sum("cantidad"))
        .order_by("sucursal_id", "-total_units", "receta__nombre")
    )
    dominant_recipe_map: dict[int, dict[str, object]] = {}
    for recipe_row in branch_recipe_rows:
        sucursal_id = int(recipe_row.get("sucursal_id") or 0)
        receta_id = int(recipe_row.get("receta_id") or 0)
        if not sucursal_id or not receta_id or sucursal_id in dominant_recipe_map:
            continue
        dominant_recipe_map[sucursal_id] = {
            "recipe_name": recipe_row.get("receta__nombre") or "Producto",
            "recipe_units": _to_decimal_safe(recipe_row.get("total_units")),
            "plan_qty": recipe_plan_map.get(receta_id, Decimal("0")),
        }

    rows: list[dict[str, object]] = []
    for row in sales_rows:
        sucursal_id = int(row.get("sucursal_id") or 0)
        if not sucursal_id:
            continue
        historico_units = _to_decimal_safe(row.get("total_units"))
        solicitud = request_map.get(sucursal_id) or {}
        solicitud_total = _to_decimal_safe(solicitud.get("solicitud_total"))
        solicitud_rows = int(solicitud.get("solicitud_rows") or 0)
        recipe_count = int(row.get("recipe_count") or 0)
        dominant_recipe = dominant_recipe_map.get(sucursal_id) or {}
        dominant_recipe_name = str(dominant_recipe.get("recipe_name") or "Producto")
        dominant_recipe_units = _to_decimal_safe(dominant_recipe.get("recipe_units"))
        dominant_plan_qty = _to_decimal_safe(dominant_recipe.get("plan_qty"))

        if solicitud_total > 0 and historico_units <= 0:
            tone = "danger"
            status = "Solicitud sin base"
            detail = "La sucursal ya trae pedido, pero el plan no tiene respaldo comparable para ese periodo."
            priority_score = solicitud_total * Decimal("10")
            action_label = "Revisar solicitud"
        elif solicitud_total > historico_units and historico_units > 0:
            tone = "warning"
            status = "Presión superior al histórico"
            detail = (
                f"Solicitud { _quantize_qty(solicitud_total) } vs histórico comparable "
                f"{ _quantize_qty(historico_units) }."
            )
            priority_score = (solicitud_total * Decimal("8")) + historico_units
            action_label = "Alinear demanda"
        elif solicitud_total > 0:
            tone = "primary"
            status = "Solicitud activa"
            detail = (
                f"La sucursal ya empuja { _quantize_qty(solicitud_total) } unidades para "
                f"{recipe_count} producto(s) del plan."
            )
            priority_score = (solicitud_total * Decimal("6")) + historico_units
            action_label = "Abrir demanda"
        else:
            tone = "success"
            status = "Demanda comparable"
            detail = (
                f"Trae { _quantize_qty(historico_units) } unidades comparables para "
                f"{recipe_count} producto(s) del plan."
            )
            priority_score = historico_units
            action_label = "Ver histórico"

        rows.append(
            {
                "sucursal_codigo": row.get("sucursal__codigo") or "",
                "sucursal_nombre": row.get("sucursal__nombre") or "Sucursal",
                "status": status,
                "tone": tone,
                "detail": detail,
                "historico_units": _quantize_qty(historico_units),
                "solicitud_total": _quantize_qty(solicitud_total),
                "solicitud_rows": solicitud_rows,
                "recipe_count": recipe_count,
                "dominant_recipe_name": dominant_recipe_name,
                "dominant_recipe_units": _quantize_qty(dominant_recipe_units),
                "dominant_plan_qty": _quantize_qty(dominant_plan_qty),
                "action_url": f"{reverse('recetas:plan_produccion')}?plan_id={plan_actual.id}&periodo={periodo}#plan-productos",
                "action_label": action_label,
                "priority_score": priority_score,
            }
        )

    tone_order = {"danger": 0, "warning": 1, "primary": 2, "success": 3}
    rows.sort(
        key=lambda item: (
            tone_order.get(str(item.get("tone") or ""), 9),
            -float(item.get("priority_score") or 0),
            str(item.get("sucursal_codigo") or ""),
        )
    )
    return rows[:6]


def _plan_branch_supply_rows(
    *,
    branch_priority_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    if not branch_priority_rows:
        return []

    recipe_name_map = {
        str(row.get("dominant_recipe_name") or ""): _quantize_qty(_to_decimal_safe(row.get("dominant_plan_qty")))
        for row in branch_priority_rows
    }
    recipe_ids = list(
        Receta.objects.filter(nombre__in=[name for name in recipe_name_map.keys() if name]).values_list("id", flat=True)
    )
    if not recipe_ids:
        return []

    lineas = list(
        LineaReceta.objects.filter(receta_id__in=recipe_ids, insumo_id__isnull=False)
        .exclude(tipo_linea=LineaReceta.TIPO_SUBSECCION)
        .select_related("insumo", "receta", "insumo__unidad_base")
    )
    if not lineas:
        return []

    canonical_by_line: dict[int, Insumo] = {}
    canonical_ids: set[int] = set()
    lineas_by_recipe_name: dict[str, list[LineaReceta]] = defaultdict(list)
    for linea in lineas:
        if not linea.insumo:
            continue
        canonical = canonical_insumo(linea.insumo) or linea.insumo
        canonical_by_line[linea.id] = canonical
        canonical_ids.add(canonical.id)
        lineas_by_recipe_name[linea.receta.nombre].append(linea)

    existencia_map = {
        int(existencia.insumo_id): existencia
        for existencia in ExistenciaInsumo.objects.filter(insumo_id__in=canonical_ids).select_related("insumo")
    }

    rows: list[dict[str, object]] = []
    for branch_row in branch_priority_rows:
        recipe_name = str(branch_row.get("dominant_recipe_name") or "")
        plan_qty = _to_decimal_safe(branch_row.get("dominant_plan_qty"))
        if not recipe_name or plan_qty <= 0:
            continue

        best_candidate: dict[str, object] | None = None
        best_score = Decimal("-1")
        for linea in lineas_by_recipe_name.get(recipe_name, []):
            canonical = canonical_by_line.get(linea.id)
            if canonical is None:
                continue
            required_qty = _to_decimal_safe(linea.cantidad) * plan_qty
            if required_qty <= 0:
                continue
            existencia = existencia_map.get(canonical.id)
            stock_actual = _to_decimal_safe(getattr(existencia, "stock_actual", 0))
            shortage = max(required_qty - stock_actual, Decimal("0"))
            readiness = _insumo_erp_readiness(canonical)
            missing = list(readiness.get("missing") or [])
            missing_cost = _latest_cost_for_insumo(canonical) is None
            score = (shortage * Decimal("100")) + (Decimal(str(len(missing))) * Decimal("50")) + required_qty
            if missing_cost:
                score += Decimal("25")
            if score > best_score:
                best_score = score
                best_candidate = {
                    "insumo_nombre": canonical.nombre,
                    "required_qty": _quantize_qty(required_qty),
                    "stock_actual": _quantize_qty(stock_actual),
                    "shortage": _quantize_qty(shortage),
                    "master_missing": missing,
                    "missing_cost": missing_cost,
                    "unidad": canonical.unidad_base.codigo if getattr(canonical, "unidad_base_id", None) else "-",
                    "action_url": reverse("maestros:insumo_update", args=[canonical.id]),
                    "action_label": "Asegurar insumo",
                }

        if best_candidate:
            rows.append(
                {
                    "sucursal_codigo": branch_row.get("sucursal_codigo") or "",
                    "sucursal_nombre": branch_row.get("sucursal_nombre") or "Sucursal",
                    "dominant_recipe_name": recipe_name,
                    **best_candidate,
                    "priority_score": best_score,
                }
            )

    rows.sort(
        key=lambda item: (
            Decimal(str(item.get("shortage") or 0)),
            Decimal(str(len(item.get("master_missing") or []))),
            Decimal(str(item.get("required_qty") or 0)),
        ),
        reverse=True,
    )
    return rows[:6]


def _plan_status_dashboard(
    plans_qs,
    *,
    start_date: date | None = None,
    end_date: date | None = None,
    group_by: str = "day",
    limit: int = 7,
) -> dict[str, Any]:
    if start_date:
        plans_qs = plans_qs.filter(fecha_produccion__gte=start_date)
    if end_date:
        plans_qs = plans_qs.filter(fecha_produccion__lte=end_date)
    if group_by not in {"day", "week", "month"}:
        group_by = "day"

    total = plans_qs.count()
    borrador = plans_qs.filter(estado=PlanProduccion.ESTADO_BORRADOR).count()
    consumo_aplicado = plans_qs.filter(estado=PlanProduccion.ESTADO_CONSUMO_APLICADO).count()
    cerrados = plans_qs.filter(estado=PlanProduccion.ESTADO_CERRADO).count()
    abiertos = borrador + consumo_aplicado
    oldest_open = plans_qs.exclude(estado=PlanProduccion.ESTADO_CERRADO).order_by("fecha_produccion", "id").first()
    oldest_open_days = 0
    if oldest_open and oldest_open.fecha_produccion:
        oldest_open_days = max((timezone.localdate() - oldest_open.fecha_produccion).days, 0)

    grouped_rows: dict[date, dict[str, Any]] = {}
    for fecha_produccion, estado in plans_qs.values_list("fecha_produccion", "estado"):
        if group_by == "week":
            bucket_date = fecha_produccion - timedelta(days=fecha_produccion.weekday())
            label = f"Semana {bucket_date.isoformat()}"
        elif group_by == "month":
            bucket_date = date(fecha_produccion.year, fecha_produccion.month, 1)
            label = bucket_date.strftime("%Y-%m")
        else:
            bucket_date = fecha_produccion
            label = bucket_date.isoformat()
        bucket = grouped_rows.setdefault(
            bucket_date,
            {
                "group_date": bucket_date,
                "label": label,
                "total": 0,
                "borrador": 0,
                "consumo_aplicado": 0,
                "cerrado": 0,
                "abiertos": 0,
            },
        )
        bucket["total"] += 1
        if estado == PlanProduccion.ESTADO_BORRADOR:
            bucket["borrador"] += 1
        elif estado == PlanProduccion.ESTADO_CONSUMO_APLICADO:
            bucket["consumo_aplicado"] += 1
        elif estado == PlanProduccion.ESTADO_CERRADO:
            bucket["cerrado"] += 1
        bucket["abiertos"] = bucket["borrador"] + bucket["consumo_aplicado"]

    rows = sorted(grouped_rows.values(), key=lambda item: item["group_date"], reverse=True)[:limit]

    if abiertos == 0:
        status = "Operacion al dia"
        tone = "success"
        detail = "No hay planes abiertos; todos los planes registrados ya quedaron cerrados."
    elif consumo_aplicado > 0:
        status = "Pendientes de cierre"
        tone = "warning"
        detail = f"Hay {consumo_aplicado} plan(es) con consumo aplicado pendientes de cierre formal."
    else:
        status = "Pendientes de ejecucion"
        tone = "warning"
        detail = f"Hay {borrador} plan(es) en borrador pendientes de ejecutar."

    return {
        "total": total,
        "borrador": borrador,
        "consumo_aplicado": consumo_aplicado,
        "cerrados": cerrados,
        "abiertos": abiertos,
        "oldest_open_name": getattr(oldest_open, "nombre", ""),
        "oldest_open_date": getattr(oldest_open, "fecha_produccion", None),
        "oldest_open_days": oldest_open_days,
        "status": status,
        "tone": tone,
        "detail": detail,
        "start_date": start_date,
        "end_date": end_date,
        "group_by": group_by,
        "group_by_label": {
            "day": "Fecha producción",
            "week": "Semana",
            "month": "Mes",
        }[group_by],
        "rows": rows,
    }


def _plan_status_dashboard_filters(request: HttpRequest) -> dict[str, Any]:
    start_date = _parse_date_safe(request.GET.get("dg_start_date"))
    end_date = _parse_date_safe(request.GET.get("dg_end_date"))
    group_by = (request.GET.get("dg_group_by") or "day").strip().lower()
    if group_by not in {"day", "week", "month"}:
        group_by = "day"
    return {
        "start_date": start_date,
        "end_date": end_date,
        "group_by": group_by,
    }


def _critical_path_rows_from_cards(
    cards: list[dict[str, object]],
    *,
    owner: str,
    fallback_url: str,
    default_dependency: str = "Inicio del flujo",
) -> list[dict[str, object]]:
    tone_completion = {
        "danger": 20,
        "warning": 55,
        "primary": 75,
        "success": 100,
    }
    stage_rows: list[dict[str, object]] = []
    for card in cards:
        tone = str(card.get("tone") or ("danger" if card.get("count") else "success"))
        count = card.get("count") or 0
        stage_rows.append(
            {
                "label": card.get("label") or "Tramo operativo",
                "owner": owner,
                "open_count": count,
                "progress_pct": tone_completion.get(tone, 0 if count else 100),
                "semaphore_tone": tone,
                "semaphore_label": "En revisión" if count else "Cerrado",
                "detail": card.get("detail") or "Sin detalle operativo",
                "depends_on": default_dependency,
                "dependency_status": card.get("detail") or "Sin dependencia registrada",
                "next_step": card.get("action_detail") or card.get("action_label") or "Revisar tramo",
                "action_label": card.get("action_label") or "Abrir",
                "action_url": card.get("action_url") or fallback_url,
            }
        )
    return _recipes_critical_path_rows(stage_rows, owner=owner, fallback_url=fallback_url)


@login_required

def recetas_list(request: HttpRequest) -> HttpResponse:
    vista = (request.GET.get("vista") or "").strip().lower()
    q = request.GET.get("q", "").strip()
    estado = request.GET.get("estado", "").strip().lower()
    health_status = request.GET.get("health_status", "").strip().lower()
    chain_status = request.GET.get("chain_status", "").strip().lower()
    chain_checkpoint = request.GET.get("chain_checkpoint", "").strip().lower()
    governance_issue = request.GET.get("governance_issue", "").strip().lower()
    enterprise_stage_filter = request.GET.get("enterprise_stage", "").strip().lower()
    tipo = request.GET.get("tipo", "").strip().upper()
    modo_operativo = request.GET.get("modo_operativo", "").strip().upper()
    familia = request.GET.get("familia", "").strip()
    categoria = request.GET.get("categoria", "").strip()
    advanced_catalog_metrics_requested = any(
        [
            health_status,
            chain_status,
            chain_checkpoint,
            governance_issue,
            enterprise_stage_filter,
            estado in {"pendientes", "ok"},
        ]
    )

    familias_db = list(
        Receta.objects.exclude(familia__exact="")
        .values_list("familia", flat=True)
        .distinct()
        .order_by("familia")
    )
    familias_catalogo = familias_producto_catalogo(familias_db)
    categorias_catalogo = list(
        Receta.objects.exclude(categoria__exact="")
        .values_list("categoria", flat=True)
        .distinct()
        .order_by("categoria")
    )
    familias_top = list(
        Receta.objects.exclude(familia__exact="")
        .values("familia")
        .annotate(total=Count("id"))
        .order_by("-total", "familia")[:8]
    )
    categorias_top = list(
        Receta.objects.exclude(categoria__exact="")
        .values("categoria")
        .annotate(total=Count("id"))
        .order_by("-total", "categoria")[:8]
    )

    recetas_base = Receta.objects.all()
    total_all = recetas_base.count()
    total_productos = recetas_base.filter(tipo=Receta.TIPO_PRODUCTO_FINAL).count()
    total_insumos = recetas_base.filter(tipo=Receta.TIPO_PREPARACION).count()
    total_subinsumos = recetas_base.filter(
        tipo=Receta.TIPO_PREPARACION,
        usa_presentaciones=True,
    ).count()
    total_batidas_base = max(total_insumos - total_subinsumos, 0)
    point_activity_snapshot = _recent_point_product_activity_snapshot()
    excluded_point_category_codes = _excluded_point_category_codes()
    product_status_rows = list(
        recetas_base.filter(tipo=Receta.TIPO_PRODUCTO_FINAL).only("id", "codigo_point", "nombre")
    )
    recetas_activas_count = sum(
        1 for receta in product_status_rows if _recipe_has_recent_point_sale(receta, point_activity_snapshot)
    )
    recetas_archivadas_count = max(total_productos - recetas_activas_count, 0)

    # Vista rápida por botones: por defecto mostramos Productos para evitar lista mezclada.
    if vista not in {"productos", "insumos", "subinsumos", "todo", "archivados"}:
        if tipo == Receta.TIPO_PRODUCTO_FINAL:
            vista = "productos"
        elif tipo == Receta.TIPO_PREPARACION:
            vista = "insumos"
        else:
            vista = "productos"

    recetas = recetas_base.select_related("rendimiento_unidad").annotate(
        pendientes_count=Count(
            "lineas",
            filter=Q(lineas__match_status=LineaReceta.STATUS_NEEDS_REVIEW),
        ),
        lineas_count=Count("lineas", distinct=True),
        presentaciones_activas_count=Count("presentaciones", filter=Q(presentaciones__activo=True)),
        has_equivalence_bom=Exists(
            RecetaEquivalencia.objects.filter(
                receta_porcion=OuterRef("pk"),
                activo=True,
            ).exclude(receta_padre=OuterRef("pk"))
        ),
        has_derived_bom=Exists(
            RecetaPresentacionDerivada.objects.filter(activo=True).filter(
                Q(receta_derivada=OuterRef("pk")) | Q(codigo_point_derivado=OuterRef("codigo_point"))
            )
        ),
    )
    if vista == "productos":
        recetas = recetas.filter(tipo=Receta.TIPO_PRODUCTO_FINAL)
    elif vista == "archivados":
        recetas = recetas.filter(tipo=Receta.TIPO_PRODUCTO_FINAL)
    elif vista == "insumos":
        recetas = recetas.filter(tipo=Receta.TIPO_PREPARACION)
    elif vista == "subinsumos":
        recetas = recetas.filter(tipo=Receta.TIPO_PREPARACION, usa_presentaciones=True)

    if modo_operativo == "BASE":
        if isinstance(recetas, list):
            recetas = [r for r in recetas if r.tipo == Receta.TIPO_PREPARACION and not r.usa_presentaciones]
        else:
            recetas = recetas.filter(tipo=Receta.TIPO_PREPARACION, usa_presentaciones=False)
    elif modo_operativo == "BASE_DERIVADOS":
        if isinstance(recetas, list):
            recetas = [r for r in recetas if r.tipo == Receta.TIPO_PREPARACION and r.usa_presentaciones]
        else:
            recetas = recetas.filter(tipo=Receta.TIPO_PREPARACION, usa_presentaciones=True)
    elif modo_operativo == "FINAL":
        if isinstance(recetas, list):
            recetas = [r for r in recetas if r.tipo == Receta.TIPO_PRODUCTO_FINAL]
        else:
            recetas = recetas.filter(tipo=Receta.TIPO_PRODUCTO_FINAL)

    if q:
        query_norm = normalizar_nombre(q)
        recetas_ranked: list[tuple[float, Receta]] = []
        for receta in recetas:
            score = _recipe_search_score(query_norm, receta)
            if score >= 60.0:
                recetas_ranked.append((score, receta))
        recetas_ranked.sort(key=lambda item: (-item[0], normalizar_nombre(item[1].nombre)))
        recetas = [item[1] for item in recetas_ranked]
    if tipo in {Receta.TIPO_PREPARACION, Receta.TIPO_PRODUCTO_FINAL}:
        if isinstance(recetas, list):
            recetas = [r for r in recetas if r.tipo == tipo]
        else:
            recetas = recetas.filter(tipo=tipo)
    if familia:
        if isinstance(recetas, list):
            recetas = [r for r in recetas if (r.familia or "") == familia]
        else:
            recetas = recetas.filter(familia=familia)
    if categoria:
        if isinstance(recetas, list):
            recetas = [r for r in recetas if (r.categoria or "") == categoria]
        else:
            recetas = recetas.filter(categoria=categoria)
    if not isinstance(recetas, list):
        if advanced_catalog_metrics_requested:
            recetas = list(
                recetas.prefetch_related("lineas", "lineas__insumo").order_by("familia", "nombre")
            )
        else:
            recetas = list(recetas.order_by("familia", "nombre"))
    for receta in recetas:
        receta.point_active = _recipe_has_recent_point_sale(receta, point_activity_snapshot)
    if vista == "archivados":
        recetas = [r for r in recetas if r.tipo == Receta.TIPO_PRODUCTO_FINAL and not r.point_active]
    elif vista == "productos":
        recetas = [r for r in recetas if r.tipo == Receta.TIPO_PRODUCTO_FINAL and r.point_active]
    if estado == "pendientes":
        recetas = [r for r in recetas if _recipe_counts_as_bom_pending(r, excluded_point_category_codes)]
    elif estado == "ok":
        recetas = [r for r in recetas if not _recipe_counts_as_bom_pending(r, excluded_point_category_codes)]

    if health_status in {"listas", "pendientes", "incompletas"}:
        recetas = [
            r
            for r in recetas
            if (health_status != "pendientes" or _recipe_counts_in_operational_catalog(r))
            and _matches_recipe_health_filter(r, health_status)
        ]
    if chain_status in {"listas", "pendientes", "incompletas"}:
        recetas = [
            r for r in recetas
            if (
                (chain_status != "pendientes" or _recipe_counts_in_operational_catalog(r))
                and (
                    (_recipe_chain_status(r)["code"] == "success" and chain_status == "listas")
                    or (_recipe_chain_status(r)["code"] == "warning" and chain_status == "pendientes")
                    or (_recipe_chain_status(r)["code"] == "danger" and chain_status == "incompletas")
                )
            )
        ]
    if chain_checkpoint in {"base_ready", "derived_sync", "final_usage", "upstream_trace", "packaging_ready", "internal_components"}:
        if isinstance(recetas, list):
            recetas = [r for r in recetas if chain_checkpoint in _recipe_failing_chain_checkpoint_codes(r)]
        else:
            recetas = [r for r in recetas if chain_checkpoint in _recipe_failing_chain_checkpoint_codes(r)]
    if governance_issue in {"familia", "categoria", "maestro_incompleto", "rendimiento", "derivados", "sync_derivados", "sin_consumo_final", "sin_base_origen", "sin_empaque", "base_directa", "componentes"}:
        if isinstance(recetas, list):
            recetas = [r for r in recetas if governance_issue in _recipe_governance_issues(r)]
        else:
            recetas = [r for r in recetas if governance_issue in _recipe_governance_issues(r)]
    if enterprise_stage_filter in {
        "base_setup",
        "derivados_setup",
        "ready_for_final",
        "feeding_final",
        "base_operativa",
        "final_setup",
        "final_normalization",
        "final_packaging",
        "final_ready",
        "final_defined",
    }:
        recetas = [r for r in recetas if _matches_recipe_enterprise_stage(r, enterprise_stage_filter)]

    total_recetas = len(recetas)
    total_pendientes = sum(1 for r in recetas if _recipe_counts_as_bom_pending(r, excluded_point_category_codes))
    total_lineas = sum((r.lineas_count or 0) for r in recetas)

    qs_filters = {
        "vista": vista,
        "q": q,
        "estado": estado,
        "health_status": health_status,
        "chain_status": chain_status,
        "chain_checkpoint": chain_checkpoint,
        "governance_issue": governance_issue,
        "enterprise_stage": enterprise_stage_filter,
        "tipo": tipo,
        "modo_operativo": modo_operativo,
        "familia": familia,
        "categoria": categoria,
    }
    qs_base = urlencode({k: v for k, v in qs_filters.items() if v})

    paginator = Paginator(recetas, 50)
    page = paginator.get_page(request.GET.get("page"))
    page_receta_ids = [receta.id for receta in page.object_list]
    page_point_validation_snapshot = (
        _build_recipe_point_validation_snapshot(list(page.object_list))
        if advanced_catalog_metrics_requested
        else {}
    )
    equivalence_by_receta_id = {
        equivalence.receta_porcion_id: equivalence
        for equivalence in RecetaEquivalencia.objects.select_related("receta_padre").filter(
            receta_porcion_id__in=page_receta_ids,
            activo=True,
        )
        if equivalence.receta_padre_id != equivalence.receta_porcion_id
    }
    derived_by_receta_id = {
        relation.receta_derivada_id: relation
        for relation in RecetaPresentacionDerivada.objects.select_related("receta_padre").filter(
            receta_derivada_id__in=page_receta_ids,
            activo=True,
        )
    }
    for receta in page.object_list:
        setattr(receta, "_effective_equivalence_cache", equivalence_by_receta_id.get(receta.id))
        setattr(receta, "_effective_derived_cache", derived_by_receta_id.get(receta.id))
        receta.costo_efectivo = _recipe_effective_cost_display(receta)
        receta.fuente_display = _recipe_source_display(receta)
        receta.bom_pending = _recipe_counts_as_bom_pending(receta, excluded_point_category_codes)
        if advanced_catalog_metrics_requested:
            receta.operational_health = _recipe_operational_health(receta)
            receta.derived_state = _recipe_derived_sync_state(receta) if receta.tipo == Receta.TIPO_PREPARACION else None
        else:
            has_pending = int(getattr(receta, "pendientes_count", 0) or 0) > 0
            receta.operational_health = {
                "code": "warning" if has_pending else "success",
                "label": "Por validar" if has_pending else "Lista para operar",
                "description": "Resumen ligero para listado.",
            }
            receta.derived_state = None
        receta.supply_chain_snapshot = None
        receta.product_upstream_snapshot = None
        receta.direct_base_snapshot = {
            "count": 0,
            "base_names": [],
            "suggested_count": 0,
            "exact_count": 0,
            "sample_suggestions": [],
        }
        if advanced_catalog_metrics_requested and receta.tipo == Receta.TIPO_PRODUCTO_FINAL:
            receta.supply_chain_snapshot = _recipe_supply_chain_snapshot(receta)
            receta.product_upstream_snapshot = getattr(receta, "_product_upstream_snapshot_cache", None)
            if receta.product_upstream_snapshot is None:
                lineas_qs = list(
                    receta.lineas.select_related("insumo").only(
                        "id",
                        "insumo_id",
                        "insumo__id",
                        "insumo__nombre",
                        "insumo__tipo_item",
                        "insumo__codigo",
                    )
                )
                receta.product_upstream_snapshot = _product_upstream_snapshot(lineas_qs, receta=receta)
                setattr(receta, "_product_upstream_snapshot_cache", receta.product_upstream_snapshot)
            receta.direct_base_snapshot = _recipe_direct_base_snapshot(receta)
        receta.governance_issues = _recipe_governance_issues(receta) if advanced_catalog_metrics_requested else []
        receta.master_gap_summary = (
            _recipe_master_gap_summary(receta)
            if advanced_catalog_metrics_requested
            else {"counts": {"unidad": 0, "proveedor": 0, "categoria": 0, "codigo_point": 0, "inactivo": 0}, "total": 0}
        )
        receta.primary_action = _recipe_primary_action(receta) if advanced_catalog_metrics_requested else None
        if receta.tipo == Receta.TIPO_PRODUCTO_FINAL:
            receta.point_validation_status = page_point_validation_snapshot.get(
                receta.id,
                {
                    "is_candidate": True,
                    "label": "Vigente operativo",
                    "description": "El producto tiene actividad reciente dentro del catálogo operativo.",
                },
            )
    health_summary = {"listas": 0, "pendientes": 0, "incompletas": 0}
    chain_summary = {"listas": 0, "pendientes": 0, "incompletas": 0}
    checkpoint_summary = {
        "base_ready": 0,
        "derived_sync": 0,
        "final_usage": 0,
        "upstream_trace": 0,
        "packaging_ready": 0,
        "internal_components": 0,
    }
    source_for_health = [
        receta for receta in (recetas if advanced_catalog_metrics_requested else page.object_list)
        if _recipe_counts_in_operational_catalog(receta)
    ]
    governance_summary = {
        "familia": 0,
        "categoria": 0,
        "maestro_incompleto": 0,
        "rendimiento": 0,
        "derivados": 0,
        "sync_derivados": 0,
        "sin_consumo_final": 0,
        "sin_base_origen": 0,
        "sin_empaque": 0,
        "base_directa": 0,
        "componentes": 0,
    }
    master_gap_totals = {
        "unidad": 0,
        "proveedor": 0,
        "categoria": 0,
        "codigo_point": 0,
        "inactivo": 0,
    }
    direct_base_suggested_total = 0
    enterprise_stage_summary = {
        "base_setup": 0,
        "derivados_setup": 0,
        "ready_for_final": 0,
        "feeding_final": 0,
        "base_operativa": 0,
        "final_setup": 0,
        "final_normalization": 0,
        "final_packaging": 0,
        "final_ready": 0,
        "final_defined": 0,
    }
    for receta in source_for_health:
        health = getattr(receta, "operational_health", None) or _recipe_operational_health(receta)
        if health["code"] == "success":
            health_summary["listas"] += 1
        elif health["code"] == "warning":
            health_summary["pendientes"] += 1
        else:
            health_summary["incompletas"] += 1
        if not advanced_catalog_metrics_requested:
            continue
        issues = getattr(receta, "governance_issues", None) or _recipe_governance_issues(receta)
        chain = getattr(receta, "chain_status_info", None) or _recipe_chain_status(receta)
        if chain["code"] == "success":
            chain_summary["listas"] += 1
        elif chain["code"] == "warning":
            chain_summary["pendientes"] += 1
        else:
            chain_summary["incompletas"] += 1
        for checkpoint_code in _recipe_failing_chain_checkpoint_codes(receta):
            if checkpoint_code in checkpoint_summary:
                checkpoint_summary[checkpoint_code] += 1
        for issue in issues:
            governance_summary[issue] += 1
        gap_summary = getattr(receta, "master_gap_summary", None) or _recipe_master_gap_summary(receta)
        for gap_key, gap_count in gap_summary["counts"].items():
            if gap_key in master_gap_totals:
                master_gap_totals[gap_key] += int(gap_count or 0)
        if receta.tipo == Receta.TIPO_PRODUCTO_FINAL:
            direct_base_snapshot = getattr(receta, "direct_base_snapshot", None) or _recipe_direct_base_snapshot(receta)
            direct_base_suggested_total += int(direct_base_snapshot.get("suggested_count") or 0)
        stage = getattr(receta, "enterprise_stage", None) or _recipe_enterprise_stage(receta)
        enterprise_stage_summary[stage["code"]] += 1
    chain_focus = _build_recipe_chain_focus(
        vista=vista,
        chain_summary=chain_summary,
        checkpoint_summary=checkpoint_summary,
        governance_summary=governance_summary,
        filters=qs_filters,
    )
    erp_governance_rows = _recipe_catalog_governance_rows(
        total_pendientes=total_pendientes,
        health_summary=health_summary,
        master_gap_totals=master_gap_totals,
        chain_focus=chain_focus,
    )
    catalog_master_blockers = int(sum(master_gap_totals.values()))
    catalog_chain_blockers = int(chain_summary["pendientes"] + chain_summary["incompletas"])
    catalog_bom_blockers = int(health_summary["pendientes"] + health_summary["incompletas"])
    downstream_handoff_rows = [
        {
            "label": "MRP",
            "owner": "Planeación / Producción",
            "status": "Listo"
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and catalog_bom_blockers == 0
            else "Bloqueado",
            "tone": "success"
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and catalog_bom_blockers == 0
            else "warning",
            "blockers": catalog_master_blockers + catalog_chain_blockers + catalog_bom_blockers,
            "completion": 100
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and catalog_bom_blockers == 0
            else max(0, 100 - ((catalog_master_blockers + catalog_chain_blockers + catalog_bom_blockers) * 3)),
            "depends_on": "Maestro + BOM + cadena cerrados",
            "exit_criteria": "El catálogo debe explotar demanda sin referencias abiertas ni documentos incompletos.",
            "detail": (
                "El catálogo ya puede alimentar el cálculo de planeación y explosión de demanda."
                if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and catalog_bom_blockers == 0
                else "MRP sigue condicionado por artículos incompletos, cadena abierta o documentos en revisión."
            ),
            "next_step": "Abrir MRP" if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and catalog_bom_blockers == 0 else "Cerrar brechas del catálogo",
            "url": reverse("recetas:mrp_form") if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and catalog_bom_blockers == 0 else (chain_focus.get("action_url") or (reverse("recetas:recetas_list") + (f'?{qs_base}' if qs_base else ""))),
            "cta": "Abrir MRP" if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and catalog_bom_blockers == 0 else (chain_focus.get("action_label") or "Abrir catálogo"),
        },
        {
            "label": "Compras",
            "owner": "Compras / Planeación",
            "status": "Listo"
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
            else "Bloqueado",
            "tone": "success"
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
            else "warning",
            "blockers": catalog_master_blockers + catalog_chain_blockers + int(governance_summary["sin_empaque"]),
            "completion": 100
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
            else max(0, 100 - ((catalog_master_blockers + catalog_chain_blockers + int(governance_summary["sin_empaque"])) * 3)),
            "depends_on": "Empaque + cadena documental",
            "exit_criteria": "Solicitudes y órdenes deben operar con productos finales completos y empaques definidos.",
            "detail": (
                "Compras ya puede usar el catálogo como referencia documental de abastecimiento."
                if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
                else "Compras sigue bloqueado por empaques faltantes, cadena abierta o artículos sin cierre maestro."
            ),
            "next_step": "Abrir compras" if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0 else "Cerrar empaque y maestro",
            "url": reverse("compras:solicitudes") if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0 else (reverse("recetas:recetas_list") + "?vista=productos&governance_issue=sin_empaque"),
            "cta": "Abrir compras" if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0 else "Ver productos finales",
        },
        {
            "label": "Inventario",
            "owner": "Inventario / Almacén",
            "status": "Listo"
            if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
            else "Bloqueado",
            "tone": "success"
            if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
            else "warning",
            "blockers": catalog_master_blockers + int(checkpoint_summary["packaging_ready"]),
            "completion": 100
            if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
            else max(0, 100 - ((catalog_master_blockers + int(checkpoint_summary["packaging_ready"])) * 4)),
            "depends_on": "Artículo maestro + empaque",
            "exit_criteria": "Inventario debe recibir productos y empaques como artículos operativos estables.",
            "detail": (
                "Inventario ya puede trabajar el catálogo sin artículos abiertos del maestro."
                if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
                else "Inventario todavía requiere artículos completos del maestro o empaques cerrados."
            ),
            "next_step": "Abrir inventario" if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0 else "Cerrar maestro e inventario",
            "url": reverse("inventario:existencias") if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0 else reverse("maestros:insumo_list"),
            "cta": "Abrir inventario" if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0 else "Abrir maestro",
        },
    ]
    trunk_handoff_rows = [
        {
            "label": "Recetas / BOM",
            "owner": "Recetas / Producción",
            "status": "Listo para operar"
            if catalog_master_blockers == 0 and catalog_bom_blockers == 0 and catalog_chain_blockers == 0
            else "Bloqueado",
            "tone": "success"
            if catalog_master_blockers == 0 and catalog_bom_blockers == 0 and catalog_chain_blockers == 0
            else "warning",
            "blockers": catalog_master_blockers + catalog_bom_blockers + catalog_chain_blockers,
            "completion": 100
            if catalog_master_blockers == 0 and catalog_bom_blockers == 0 and catalog_chain_blockers == 0
            else max(0, 100 - ((catalog_master_blockers + catalog_bom_blockers + catalog_chain_blockers) * 3)),
            "depends_on": "Maestro cerrado + estructura base-derivados-final consistente",
            "exit_criteria": "Las recetas deben costear, normalizar y alimentar producto final sin referencias abiertas.",
            "detail": (
                "El catálogo ya puede operar como documento BOM estable."
                if catalog_master_blockers == 0 and catalog_bom_blockers == 0 and catalog_chain_blockers == 0
                else "Todavía hay brechas en maestro, cadena o estructura BOM."
            ),
            "next_step": chain_focus.get("action_detail") or "Cerrar catálogo de recetas",
            "url": chain_focus.get("action_url") or (reverse("recetas:recetas_list") + (f'?{qs_base}' if qs_base else "")),
            "cta": chain_focus.get("action_label") or "Abrir recetas",
        },
        {
            "label": "Compras documentales",
            "owner": "Compras / Planeación",
            "status": "Listo para operar"
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
            else "Bloqueado",
            "tone": "success"
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
            else "warning",
            "blockers": catalog_master_blockers + catalog_chain_blockers + int(governance_summary["sin_empaque"]),
            "completion": 100
            if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
            else max(0, 100 - ((catalog_master_blockers + catalog_chain_blockers + int(governance_summary["sin_empaque"])) * 3)),
            "depends_on": "Producto final completo + empaques definidos",
            "exit_criteria": "Solicitudes, órdenes y recepciones deben usar productos finales cerrados documentalmente.",
            "detail": (
                "Compras ya puede tomar el catálogo como documento operativo de abastecimiento."
                if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0
                else "Compras sigue condicionado por maestro, cadena o empaques pendientes."
            ),
            "next_step": "Abrir compras" if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0 else "Cerrar empaque y maestro",
            "url": reverse("compras:solicitudes") if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0 else (reverse("recetas:recetas_list") + "?vista=productos&governance_issue=sin_empaque"),
            "cta": "Abrir compras" if catalog_master_blockers == 0 and catalog_chain_blockers == 0 and governance_summary["sin_empaque"] == 0 else "Ver productos finales",
        },
        {
            "label": "Inventario / Reabasto",
            "owner": "Inventario / Almacén",
            "status": "Listo para operar"
            if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
            else "Bloqueado",
            "tone": "success"
            if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
            else "warning",
            "blockers": catalog_master_blockers + int(checkpoint_summary["packaging_ready"]),
            "completion": 100
            if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
            else max(0, 100 - ((catalog_master_blockers + int(checkpoint_summary["packaging_ready"])) * 4)),
            "depends_on": "Artículos listos + empaques operativos + cierre documental",
            "exit_criteria": "Inventario debe trabajar finales y empaques como artículos estables de stock y reabasto.",
            "detail": (
                "Inventario ya puede sostener existencia y reabasto sobre el catálogo final."
                if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0
                else "Inventario todavía requiere cierre de maestro o empaque para operar estable."
            ),
            "next_step": "Abrir inventario" if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0 else "Cerrar maestro e inventario",
            "url": reverse("inventario:existencias") if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0 else reverse("maestros:insumo_list"),
            "cta": "Abrir inventario" if catalog_master_blockers == 0 and checkpoint_summary["packaging_ready"] == 0 else "Abrir maestro",
        },
    ]
    if advanced_catalog_metrics_requested:
        for receta in page.object_list:
            receta.chain_status_info = _recipe_chain_status(receta)
            receta.chain_checkpoints = _recipe_chain_checkpoints(receta)
            receta.chain_action_links = _recipe_chain_actions_catalog(receta)
            receta.chain_focus_summary = _recipe_chain_focus_summary(receta)
            receta.enterprise_stage = _recipe_enterprise_stage(receta)
            receta.enterprise_stage_playbook = _recipe_enterprise_stage_playbook(receta)
            receta.enterprise_stage_progress = _recipe_stage_progress(receta.enterprise_stage_playbook)
            receta.document_status = _recipe_document_status(receta)
    point_recipe_sync_panel = _point_recipe_sync_job_panel(request)
    return render(
        request,
        "recetas/recetas_list.html",
        {
            "page": page,
            "vista": vista,
            "vista_actual": vista,
            "q": q,
            "estado": estado,
            "health_status": health_status,
            "chain_status": chain_status,
            "chain_checkpoint": chain_checkpoint,
            "governance_issue": governance_issue,
            "tipo": tipo,
            "modo_operativo": modo_operativo,
            "familia": familia,
            "categoria": categoria,
            "familias_catalogo": familias_catalogo,
            "categorias_catalogo": categorias_catalogo,
            "familias_top": familias_top,
            "categorias_top": categorias_top,
            "total_recetas": total_recetas,
            "total_pendientes": total_pendientes,
            "total_lineas": total_lineas,
            "total_all": total_all,
            "total_productos": total_productos,
            "total_insumos": total_insumos,
            "total_batidas_base": total_batidas_base,
            "total_subinsumos": total_subinsumos,
            "recetas_activas_count": recetas_activas_count,
            "recetas_archivadas_count": recetas_archivadas_count,
            "health_summary": health_summary,
            "chain_summary": chain_summary,
            "chain_focus": chain_focus,
            "enterprise_stage_filter": enterprise_stage_filter,
            "enterprise_stage_summary": enterprise_stage_summary,
            "checkpoint_summary": checkpoint_summary,
            "governance_summary": governance_summary,
            "master_gap_totals": master_gap_totals,
            "erp_command_center": {
                "owner": "Recetas / Producción",
                "status": "Crítico"
                if (health_summary["pendientes"] + health_summary["incompletas"] + sum(master_gap_totals.values())) > 0
                else "Controlado",
                "tone": "danger"
                if (health_summary["pendientes"] + health_summary["incompletas"] + sum(master_gap_totals.values())) > 0
                else "success",
                "blockers": health_summary["pendientes"] + health_summary["incompletas"] + sum(master_gap_totals.values()),
                "next_step": chain_focus.get("action_detail") or "Continuar el cierre operativo del catálogo de recetas.",
                "url": chain_focus.get("action_url") or qs_base,
                "cta": chain_focus.get("action_label") or "Abrir catálogo",
            },
            "erp_governance_rows": erp_governance_rows,
            "downstream_handoff_rows": downstream_handoff_rows,
            "trunk_handoff_rows": trunk_handoff_rows,
            "trunk_handoff_summary": _trunk_handoff_summary(
                trunk_handoff_rows,
                owner="Recetas / Producción",
                fallback_url=qs_base,
            ),
            "executive_radar_rows": _recipes_executive_radar_rows(
                erp_governance_rows,
                owner="Recetas / Producción",
                fallback_url=qs_base,
            ),
            "critical_path_rows": _recipes_critical_path_rows(
                erp_governance_rows,
                owner="Recetas / Producción",
                fallback_url=qs_base,
            ),
            "direct_base_suggested_total": direct_base_suggested_total,
            "qs_base": qs_base,
            "point_recipe_sync_panel": point_recipe_sync_panel,
        },
    )


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def recetas_sync_all(request: HttpRequest) -> HttpResponse:
    next_url = request.POST.get("next") or reverse("recetas:recetas_list")
    try:
        job, summary = _run_point_recipe_sync_action(request, action_label="SYNC_ALL_RECIPES")
        snapshot = _snapshot_costeo_after_sync()
        messages.success(
            request,
            (
                f"{_format_point_recipe_sync_message(summary)} "
                f"{summary.get('lineas_created', 0)} líneas materializadas. "
                f"Corte semanal {snapshot['week_start']} recalculado."
            ),
        )
        log_event(
            request.user,
            "SNAPSHOT_WEEKLY_COST",
            "recetas.RecetaCostoSemanal",
            None,
            {"source": "SYNC_ALL_RECIPES", "job_id": job.id, "snapshot": snapshot},
        )
    except Exception as exc:
        messages.error(request, f"No se pudo actualizar recetas desde Point: {exc}")
    return redirect(next_url)


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def recetas_sync_group(request: HttpRequest) -> HttpResponse:
    next_url = request.POST.get("next") or reverse("recetas:recetas_list")
    qs = _filtered_recipe_sync_queryset(request.POST).exclude(codigo_point__exact="")
    product_codes = list(qs.values_list("codigo_point", flat=True).distinct())
    if not product_codes:
        messages.warning(request, "No hay recetas filtradas con código Point para sincronizar.")
        return redirect(next_url)
    try:
        job, summary = _run_point_recipe_sync_action(
            request,
            action_label="SYNC_FILTERED_GROUP",
            product_codes=product_codes,
        )
        receta_ids = list(Receta.objects.filter(codigo_point__in=product_codes).values_list("id", flat=True))
        snapshot = _snapshot_costeo_after_sync(receta_ids=receta_ids)
        messages.success(
            request,
            (
                f"Grupo actualizado desde Point: {len(product_codes)} códigos enviados. "
                f"{_format_point_recipe_sync_message(summary)} "
                f"Snapshot semanal del grupo recalculado."
            ),
        )
        log_event(
            request.user,
            "SNAPSHOT_WEEKLY_COST",
            "recetas.RecetaCostoSemanal",
            None,
            {"source": "SYNC_FILTERED_GROUP", "job_id": job.id, "snapshot": snapshot, "product_codes": product_codes},
        )
    except Exception as exc:
        messages.error(request, f"No se pudo actualizar el grupo filtrado: {exc}")
    return redirect(next_url)


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def recetas_sync_new(request: HttpRequest) -> HttpResponse:
    next_url = request.POST.get("next") or reverse("recetas:recetas_list")
    service = PointProductRecipeSyncService()
    try:
        discovery = service.discover_new_product_codes(branch_hint="MATRIZ")
        new_codes = discovery.get("new_codes") or []
        blocked_candidates_count = int(discovery.get("blocked_candidates_count") or 0)
        if not new_codes:
            if blocked_candidates_count:
                messages.warning(request, _format_point_recipe_discovery_blocked_message(discovery))
            else:
                messages.info(request, "Point no reportó productos nuevos con receta pendientes de incorporar.")
            return redirect(next_url)
        job, summary = _run_point_recipe_sync_action(
            request,
            action_label="SYNC_ONLY_NEW_PRODUCTS",
            product_codes=list(new_codes),
        )
        receta_ids = list(Receta.objects.filter(codigo_point__in=new_codes).values_list("id", flat=True))
        snapshot = _snapshot_costeo_after_sync(receta_ids=receta_ids)
        messages.success(
            request,
            (
                f"{_format_point_recipe_sync_message(summary, new_codes_count=len(new_codes))} "
                f"Snapshot semanal actualizado."
            ),
        )
        if blocked_candidates_count:
            messages.warning(request, _format_point_recipe_discovery_blocked_message(discovery))
        log_event(
            request.user,
            "SYNC_ONLY_NEW_PRODUCTS",
            "pos_bridge.PointSyncJob",
            job.id,
            {"discovery": discovery, "snapshot": snapshot},
        )
    except Exception as exc:
        messages.error(request, f"No se pudo incorporar productos nuevos desde Point: {exc}")
    return redirect(next_url)


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def receta_sync_point(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    next_url = request.POST.get("next") or reverse("recetas:receta_detail", args=[receta.id])
    if not (receta.codigo_point or "").strip():
        messages.warning(request, "La receta no tiene código Point para sincronizar.")
        return redirect(next_url)
    try:
        job, summary = _run_point_recipe_sync_action(
            request,
            action_label="SYNC_SINGLE_RECIPE",
            product_codes=[receta.codigo_point],
        )
        snapshot = _snapshot_costeo_after_sync(receta_ids=[receta.id])
        messages.success(
            request,
            (
                f"{_format_point_recipe_sync_message(summary)} "
                f"{summary.get('lineas_created', 0)} líneas materializadas y costeo semanal recalculado."
            ),
        )
        log_event(
            request.user,
            "SNAPSHOT_WEEKLY_COST",
            "recetas.RecetaCostoSemanal",
            None,
            {"source": "SYNC_SINGLE_RECIPE", "job_id": job.id, "receta_id": receta.id, "snapshot": snapshot},
        )
    except Exception as exc:
        messages.error(request, f"No se pudo actualizar la receta desde Point: {exc}")
    return redirect(next_url)


def _costeo_scope_queryset(selected_week: date, *, scope: str, familia: str = "", temporalidad: str = "", q: str = ""):
    qs = RecetaCostoSemanal.objects.filter(week_start=selected_week).select_related("receta", "base_receta", "addon_receta", "addon_rule")
    if scope == "productos":
        qs = qs.filter(scope_type=RecetaCostoSemanal.SCOPE_RECIPE, receta__tipo=Receta.TIPO_PRODUCTO_FINAL)
    elif scope == "bases":
        qs = qs.filter(scope_type=RecetaCostoSemanal.SCOPE_RECIPE, receta__tipo=Receta.TIPO_PREPARACION)
    elif scope == "addons":
        qs = qs.filter(scope_type=RecetaCostoSemanal.SCOPE_GROUPED_ADDON)
    if familia:
        qs = qs.filter(familia=familia)
    if temporalidad:
        qs = qs.filter(temporalidad=temporalidad)
    if q:
        qs = qs.filter(
            Q(label__icontains=q)
            | Q(base_receta__nombre__icontains=q)
            | Q(addon_receta__nombre__icontains=q)
            | Q(receta__nombre__icontains=q)
        )
    return qs.order_by("label", "id")


@login_required
@permission_required("recetas.view_receta", raise_exception=True)
def costeo_dashboard(request: HttpRequest) -> HttpResponse:
    week_values = list(
        RecetaCostoSemanal.objects.order_by("-week_start").values_list("week_start", flat=True).distinct()[:26]
    )
    selected_week = None
    week_raw = (request.GET.get("week_start") or "").strip()
    if week_raw:
        try:
            selected_week = date.fromisoformat(week_raw)
        except ValueError:
            selected_week = None
    if selected_week is None and week_values:
        selected_week = week_values[0]

    scope = (request.GET.get("scope") or "productos").strip().lower()
    if scope not in {"productos", "bases", "addons", "todo"}:
        scope = "productos"
    familia = (request.GET.get("familia") or "").strip()
    temporalidad = (request.GET.get("temporalidad") or "").strip().upper()
    q = (request.GET.get("q") or "").strip()

    familias_catalogo = list(
        RecetaCostoSemanal.objects.exclude(familia__exact="")
        .values_list("familia", flat=True)
        .distinct()
        .order_by("familia")
    )

    rows = []
    selected_identity = None
    history_rows = []
    if selected_week is not None:
        rows = list(_costeo_scope_queryset(selected_week, scope=scope, familia=familia, temporalidad=temporalidad, q=q))
        selected_identity = (request.GET.get("identity") or "").strip() or (rows[0].identity_key if rows else "")
        if selected_identity:
            history_rows = list(
                RecetaCostoSemanal.objects.filter(identity_key=selected_identity)
                .select_related("receta", "base_receta", "addon_receta")
                .order_by("-week_start", "-id")[:12]
            )
    total_cost = sum(Decimal(row.costo_total or 0) for row in rows)
    total_delta = sum(Decimal(row.delta_total or 0) for row in rows if row.delta_total is not None)
    increase_rows = sorted(
        [row for row in rows if row.delta_total is not None and Decimal(row.delta_total) > 0],
        key=lambda item: Decimal(item.delta_total),
        reverse=True,
    )[:10]
    decrease_rows = sorted(
        [row for row in rows if row.delta_total is not None and Decimal(row.delta_total) < 0],
        key=lambda item: Decimal(item.delta_total),
    )[:10]
    stable_rows = len([row for row in rows if row.delta_total in (None, Decimal("0"), 0)])
    selected_row = next((row for row in rows if row.identity_key == selected_identity), None)
    max_history_cost = max((Decimal(item.costo_total or 0) for item in history_rows), default=Decimal("0"))
    for item in history_rows:
        current_cost = Decimal(item.costo_total or 0)
        item.history_pct = float((current_cost / max_history_cost) * Decimal("100")) if max_history_cost > 0 else 0.0
    previous_week = week_values[1] if len(week_values) > 1 else None

    return render(
        request,
        "recetas/costeo_dashboard.html",
        {
            "selected_week": selected_week,
            "week_values": week_values,
            "previous_week": previous_week,
            "scope": scope,
            "familia": familia,
            "temporalidad": temporalidad,
            "q": q,
            "familias_catalogo": familias_catalogo,
            "rows": rows,
            "increase_rows": increase_rows,
            "decrease_rows": decrease_rows,
            "stable_rows": stable_rows,
            "total_cost": total_cost,
            "total_delta": total_delta,
            "selected_identity": selected_identity,
            "selected_row": selected_row,
            "history_rows": history_rows,
            "max_history_cost": max_history_cost,
        },
    )


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def costeo_dashboard_snapshot(request: HttpRequest) -> HttpResponse:
    next_url = request.POST.get("next") or reverse("recetas:costeo_dashboard")
    try:
        summary = snapshot_weekly_costs()
        messages.success(
            request,
            (
                f"Corte semanal de costeo actualizado: semana {summary.week_start} a {summary.week_end}, "
                f"{summary.total_items} registros procesados."
            ),
        )
        log_event(
            request.user,
            "SNAPSHOT_WEEKLY_COST",
            "recetas.RecetaCostoSemanal",
            None,
            {
                "week_start": summary.week_start.isoformat(),
                "week_end": summary.week_end.isoformat(),
                "recipes_created": summary.recipes_created,
                "recipes_updated": summary.recipes_updated,
                "addons_created": summary.addons_created,
                "addons_updated": summary.addons_updated,
            },
        )
    except Exception as exc:
        messages.error(request, f"No se pudo generar el corte semanal de costeo: {exc}")
    return redirect(next_url)


@login_required
def monitor_margenes(request: HttpRequest) -> HttpResponse:
    if not can_view_recetas(request.user):
        raise PermissionDenied

    q = (request.GET.get("q") or "").strip()
    reference_end = timezone.localdate()
    reference_start = reference_end - timedelta(days=90)
    branch_ids = set(sucursales_operativas(reference_end).values_list("id", flat=True))

    latest_versions: list[RecetaCostoVersion] = []
    seen_recipe_ids: set[int] = set()
    version_qs = (
        RecetaCostoVersion.objects.filter(costo_total__gt=0)
        .select_related("receta")
        .order_by("receta_id", "-version_num", "-id")
    )
    for version in version_qs:
        if version.receta_id in seen_recipe_ids:
            continue
        seen_recipe_ids.add(version.receta_id)
        latest_versions.append(version)

    if q:
        normalized_q = normalizar_nombre(q)
        latest_versions = [
            version
            for version in latest_versions
            if normalized_q in normalizar_nombre(version.receta.nombre)
            or normalized_q in normalizar_nombre(version.receta.codigo_point or "")
            or normalized_q in normalizar_nombre(version.receta.familia or "")
        ]

    recipe_ids = {version.receta_id for version in latest_versions}
    price_map = resolve_unit_prices_bulk(
        recipe_ids,
        reference_start,
        reference_end,
        branch_ids=branch_ids,
    )

    rows: list[dict[str, Any]] = []
    red_count = 0
    yellow_count = 0
    green_count = 0
    missing_price_count = 0
    total_cost = Decimal("0")
    total_price = Decimal("0")

    for version in latest_versions:
        branch_prices = [
            Decimal(price)
            for (recipe_id, _branch_id), price in price_map.items()
            if recipe_id == version.receta_id and Decimal(price or 0) > 0
        ]
        avg_price = (
            (sum(branch_prices, Decimal("0")) / Decimal(len(branch_prices))).quantize(Decimal("0.01"))
            if branch_prices
            else Decimal("0")
        )
        cost = _recipe_effective_cost_display(version.receta).quantize(Decimal("0.01"))
        if cost <= 0:
            continue
        margin_pct = None
        status = "SIN_PRECIO"
        status_label = "Sin precio"
        sort_margin = Decimal("-1")
        if avg_price > 0:
            margin_pct = (((avg_price - cost) / avg_price) * Decimal("100")).quantize(Decimal("0.01"))
            sort_margin = margin_pct
            if margin_pct < Decimal("40"):
                status = "ALERTA"
                status_label = "Rojo < 40%"
                red_count += 1
            elif margin_pct < Decimal("55"):
                status = "REVISION"
                status_label = "Amarillo < 55%"
                yellow_count += 1
            else:
                status = "SALUDABLE"
                status_label = "Verde >= 55%"
                green_count += 1
            total_price += avg_price
        else:
            missing_price_count += 1
        total_cost += cost

        rows.append(
            {
                "receta": version.receta,
                "version": version,
                "price": avg_price,
                "cost": cost,
                "margin_pct": margin_pct,
                "status": status,
                "status_label": status_label,
                "price_points": len(branch_prices),
                "sort_margin": sort_margin,
                "cost_source_label": "Costo actual",
            }
        )

    rows.sort(key=lambda item: (item["sort_margin"], item["receta"].nombre.lower()))

    return render(
        request,
        "recetas/monitor_margenes.html",
        {
            "rows": rows,
            "q": q,
            "reference_start": reference_start,
            "reference_end": reference_end,
            "red_count": red_count,
            "yellow_count": yellow_count,
            "green_count": green_count,
            "missing_price_count": missing_price_count,
            "total_cost": total_cost,
            "total_price": total_price,
        },
    )


@login_required
@permission_required("recetas.add_receta", raise_exception=True)
def receta_create(request: HttpRequest) -> HttpResponse:
    familias_db = list(
        Receta.objects.exclude(familia__exact="")
        .values_list("familia", flat=True)
        .distinct()
        .order_by("familia")
    )
    familias_catalogo = familias_producto_catalogo(familias_db)
    categorias_catalogo = list(
        Receta.objects.exclude(categoria__exact="")
        .values_list("categoria", flat=True)
        .distinct()
        .order_by("categoria")
    )
    unidades = UnidadMedida.objects.order_by("codigo")

    recipe_modes = [
        {
            "code": "BASE",
            "title": "Insumo interno base",
            "description": "Batida o mezcla producida con materia prima. Puede llevar rendimiento y luego generar presentaciones.",
        },
        {
            "code": "BASE_DERIVADOS",
            "title": "Insumo base con presentaciones",
            "description": "Usa este modo si de la batida saldrán tamaños o subinsumos derivados como chico, mediano, grande o rosca.",
        },
        {
            "code": "FINAL",
            "title": "Producto final de venta",
            "description": "Artículo final armado con panes, rellenos, coberturas, empaques u otros componentes.",
        },
    ]

    defaults = {
        "nombre": "",
        "codigo_point": "",
        "familia": "",
        "categoria": "",
        "sheet_name": "",
        "recipe_mode": "BASE",
        "rendimiento_cantidad": "",
        "rendimiento_unidad_id": "",
    }

    requested_mode = (request.GET.get("mode") or "").strip().upper()
    if requested_mode in {"BASE", "BASE_DERIVADOS", "FINAL"}:
        defaults["recipe_mode"] = requested_mode
    source_base_receta = None
    source_base_context = None
    source_base_raw = (request.GET.get("source_base") or "").strip()
    if source_base_raw.isdigit():
        source_base_receta = (
            Receta.objects.filter(pk=int(source_base_raw), tipo=Receta.TIPO_PREPARACION)
            .only(
                "id",
                "nombre",
                "familia",
                "categoria",
                "sheet_name",
                "usa_presentaciones",
                "rendimiento_cantidad",
                "rendimiento_unidad_id",
            )
            .first()
        )
        if source_base_receta and defaults["recipe_mode"] == "FINAL":
            if source_base_receta.familia:
                defaults["familia"] = source_base_receta.familia
            defaults["categoria"] = source_base_receta.categoria or defaults["categoria"]
            defaults["sheet_name"] = source_base_receta.sheet_name or defaults["sheet_name"]
            derived_state = _recipe_derived_sync_state(source_base_receta)
            active_presentaciones = int(derived_state["active_presentaciones"] or 0)
            derived_presentaciones = int(derived_state["derived_presentaciones"] or 0)
            source_base_context = {
                "base_ready": bool(
                    source_base_receta.rendimiento_cantidad
                    and source_base_receta.rendimiento_unidad_id
                    and derived_state["prep_ready"]
                ),
                "uses_presentaciones": bool(source_base_receta.usa_presentaciones),
                "active_presentaciones": active_presentaciones,
                "derived_presentaciones": derived_presentaciones,
                "sync_missing_count": max(active_presentaciones - derived_presentaciones, 0),
                "next_step_label": (
                    "La base ya puede alimentar el producto final."
                    if active_presentaciones <= 0 or derived_presentaciones >= active_presentaciones
                    else "Conviene sincronizar derivados antes de armar el producto final."
                ),
                "sync_url": reverse("recetas:receta_sync_derivados", args=[source_base_receta.id]),
                "detail_url": reverse("recetas:receta_detail", args=[source_base_receta.id]),
            }

    if request.method == "POST":
        values = {
            "nombre": (request.POST.get("nombre") or "").strip(),
            "codigo_point": (request.POST.get("codigo_point") or "").strip(),
            "familia": (request.POST.get("familia") or "").strip(),
            "categoria": (request.POST.get("categoria") or "").strip(),
            "sheet_name": (request.POST.get("sheet_name") or "").strip(),
            "recipe_mode": (request.POST.get("recipe_mode") or "BASE").strip().upper(),
            "rendimiento_cantidad": (request.POST.get("rendimiento_cantidad") or "").strip(),
            "rendimiento_unidad_id": (request.POST.get("rendimiento_unidad_id") or "").strip(),
        }
        mode = values["recipe_mode"] if values["recipe_mode"] in {"BASE", "BASE_DERIVADOS", "FINAL"} else "BASE"

        if not values["nombre"]:
            messages.error(request, "El nombre de la receta es obligatorio.")
        else:
            validation_errors = _validate_receta_enterprise_rules(
                mode_code=mode,
                usa_presentaciones=(mode == "BASE_DERIVADOS"),
                rendimiento_cantidad=_to_decimal_or_none(values["rendimiento_cantidad"]),
                rendimiento_unidad_id=values["rendimiento_unidad_id"],
                familia=values["familia"],
            )
            if validation_errors:
                for error in validation_errors:
                    messages.error(request, error)
            else:
                rendimiento_cantidad = (
                    None if mode == "FINAL" else _to_decimal_or_none(values["rendimiento_cantidad"])
                )
                rendimiento_unidad = (
                    None
                    if mode == "FINAL" or not values["rendimiento_unidad_id"]
                    else UnidadMedida.objects.filter(pk=values["rendimiento_unidad_id"]).first()
                )
                receta = Receta.objects.create(
                    nombre=values["nombre"][:250],
                    codigo_point=values["codigo_point"][:80],
                    familia=values["familia"][:120],
                    categoria=values["categoria"][:120],
                    sheet_name=values["sheet_name"][:120],
                    tipo=Receta.TIPO_PRODUCTO_FINAL if mode == "FINAL" else Receta.TIPO_PREPARACION,
                    usa_presentaciones=(mode == "BASE_DERIVADOS"),
                    rendimiento_cantidad=rendimiento_cantidad,
                    rendimiento_unidad=rendimiento_unidad,
                    hash_contenido=uuid4().hex,
                )
                log_event(
                    request.user,
                    "CREATE_RECETA",
                    "recetas.Receta",
                    str(receta.pk),
                    {
                        "nombre": receta.nombre,
                        "tipo": receta.tipo,
                        "usa_presentaciones": receta.usa_presentaciones,
                        "familia": receta.familia,
                        "categoria": receta.categoria,
                    },
                )
                if mode == "FINAL":
                    messages.success(
                        request,
                        "Producto final creado. Ahora agrega sus componentes de armado.",
                    )
                elif mode == "BASE_DERIVADOS":
                    messages.success(
                        request,
                        "Insumo base creado. El siguiente paso es definir sus presentaciones derivadas.",
                    )
                else:
                    messages.success(
                        request,
                        "Insumo interno base creado. Ahora captura su receta y rendimiento.",
                    )
                return redirect("recetas:receta_detail", pk=receta.pk)

        defaults.update(values)

    return render(
        request,
        "recetas/receta_create.html",
        {
            "recipe_modes": recipe_modes,
            "familias_catalogo": familias_catalogo,
            "categorias_catalogo": categorias_catalogo,
            "familia_categoria_catalogo_json": familia_categoria_catalogo_json(categorias_catalogo),
            "unidades": unidades,
            "values": defaults,
            "source_base_receta": source_base_receta,
            "source_base_context": source_base_context,
        },
    )


@login_required
def receta_detail(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    familias_db = list(
        Receta.objects.exclude(familia__exact="")
        .values_list("familia", flat=True)
        .distinct()
        .order_by("familia")
    )
    familias_catalogo = familias_producto_catalogo(familias_db)
    categorias_catalogo = list(
        Receta.objects.exclude(categoria__exact="")
        .values_list("categoria", flat=True)
        .distinct()
        .order_by("categoria")
    )
    versiones_unavailable = False
    versiones_all = []
    versiones_recientes = []
    comparativo = None
    try:
        if not receta.versiones_costo.exists():
            _sync_cost_version_safe(request, receta, "AUTO_BOOTSTRAP")
        versiones_all = _load_versiones_costeo(receta, 80)
        versiones_recientes = versiones_all[:8]
        comparativo = comparativo_versiones(versiones_recientes)
    except (OperationalError, ProgrammingError):
        versiones_unavailable = True
        messages.warning(
            request,
            "Histórico de versiones de costo no disponible en este entorno. Ejecuta migraciones para habilitarlo.",
        )

    lineas = list(receta.lineas.select_related("insumo").order_by("posicion"))
    presentaciones = sorted(list(receta.presentaciones.all()), key=lambda p: _presentacion_sort_key(p.nombre))
    costeo_unavailable = False
    try:
        costeo_actual = calcular_costeo_receta(receta)
    except (OperationalError, ProgrammingError):
        costeo_unavailable = True
        costeo_actual = _empty_costeo_actual()
        messages.warning(
            request,
            "Costeo avanzado no disponible en este entorno. Ejecuta migraciones para habilitar drivers.",
        )

    selected_base = None
    selected_target = None
    compare_data = None
    base_raw = (request.GET.get("base") or "").strip()
    target_raw = (request.GET.get("target") or "").strip()
    try:
        if base_raw and target_raw:
            base_num = int(base_raw)
            target_num = int(target_raw)
            by_num = {v.version_num: v for v in versiones_all}
            if base_num in by_num and target_num in by_num and base_num != target_num:
                selected_base = by_num[base_num]
                selected_target = by_num[target_num]
    except Exception:
        selected_base = None
        selected_target = None

    if not selected_base and not selected_target and len(versiones_all) >= 2:
        selected_target = versiones_all[0]
        selected_base = versiones_all[1]

    if selected_base and selected_target:
        compare_data = _compare_versions(selected_base, selected_target)

    def _line_origin_bucket(linea: LineaReceta) -> str:
        if linea.tipo_linea == LineaReceta.TIPO_SUBSECCION:
            return "subsecciones"
        if not linea.insumo:
            return "pendientes"
        if linea.insumo.tipo_item == Insumo.TIPO_EMPAQUE:
            return "empaques"
        if linea.insumo.tipo_item == Insumo.TIPO_INTERNO or (linea.insumo.codigo or "").startswith("DERIVADO:RECETA:"):
            return "internos"
        return "materia_prima"

    total_lineas = len(lineas)
    total_match = sum(1 for l in lineas if l.match_status == LineaReceta.STATUS_AUTO)
    total_revision = sum(1 for l in lineas if l.match_status == LineaReceta.STATUS_NEEDS_REVIEW)
    total_sin_match = sum(1 for l in lineas if l.match_status == LineaReceta.STATUS_REJECTED)
    total_costo_directo = sum((l.costo_total_estimado or 0.0) for l in lineas)
    total_costo_estimado = float(receta.costo_total_estimado or 0.0)
    total_materia_prima = sum(1 for l in lineas if l.insumo and l.insumo.tipo_item == Insumo.TIPO_MATERIA_PRIMA)
    total_internos = sum(
        1
        for l in lineas
        if l.insumo and (l.insumo.tipo_item == Insumo.TIPO_INTERNO or (l.insumo.codigo or "").startswith("DERIVADO:RECETA:"))
    )
    total_empaques = sum(1 for l in lineas if l.insumo and l.insumo.tipo_item == Insumo.TIPO_EMPAQUE)
    total_subsecciones = sum(1 for l in lineas if l.tipo_linea == LineaReceta.TIPO_SUBSECCION)
    receta_rol_label = (
        "Producto final de venta"
        if receta.tipo == Receta.TIPO_PRODUCTO_FINAL
        else "Insumo interno / batida base"
    )
    receta_rol_description = (
        "Se arma con subinsumos, insumos internos, materia prima puntual y empaques."
        if receta.tipo == Receta.TIPO_PRODUCTO_FINAL
        else "Se produce con materia prima y puede generar subinsumos o presentaciones derivadas."
    )
    is_producto_final = receta.tipo == Receta.TIPO_PRODUCTO_FINAL
    is_base_con_presentaciones = receta.tipo == Receta.TIPO_PREPARACION and receta.usa_presentaciones
    is_base_simple = receta.tipo == Receta.TIPO_PREPARACION and not receta.usa_presentaciones
    supply_chain_snapshot = _recipe_supply_chain_snapshot(receta)
    active_presentaciones = [p for p in presentaciones if p.activo]
    presentacion_derivados_activos = Insumo.objects.filter(
        codigo__startswith=f"DERIVADO:RECETA:{receta.id}:PRESENTACION:",
        activo=True,
    ).count()
    add_line_label = "Agregar componente" if is_producto_final else "Agregar insumo base"
    copy_label = "Copiar estructura base" if is_producto_final else "Copiar ingredientes"
    chain_focus_summary = _recipe_chain_focus_summary(receta)
    enterprise_stage = _recipe_enterprise_stage(receta)
    enterprise_stage_playbook = _recipe_enterprise_stage_playbook(receta)
    next_steps = []
    chain_actions: list[dict[str, str]] = []
    base_chain_actions: list[dict[str, str]] = []
    base_chain_checkpoints: list[dict[str, object]] = []
    if is_base_simple:
        next_steps = [
            "Captura materias primas y cantidades reales de la batida.",
            "Define rendimiento para calcular costo por unidad base.",
            "Si después saldrán tamaños, activa presentaciones y crea los derivados.",
        ]
        base_chain_checkpoints = [
            {
                "label": "Rendimiento capturado",
                "ok": bool(receta.rendimiento_cantidad and receta.rendimiento_cantidad > 0),
                "detail": "Costo unitario base listo para operar." if receta.rendimiento_cantidad else "Todavía falta rendimiento total.",
            },
            {
                "label": "Unidad de rendimiento",
                "ok": bool(receta.rendimiento_unidad_id),
                "detail": receta.rendimiento_unidad.codigo if receta.rendimiento_unidad_id else "Sin unidad base definida.",
            },
        ]
    elif is_base_con_presentaciones:
        next_steps = [
            "Captura la receta completa de la batida o mezcla base.",
            "Mantén actualizado el rendimiento total para costeo por kg/lt/pza.",
            "Administra presentaciones derivadas para que aparezcan como subinsumos en productos finales.",
        ]
        prep_insumo_operativo = bool(supply_chain_snapshot and supply_chain_snapshot["prep_insumo"])
        chain_actions.append(
            {"label": "Agregar presentación", "url": reverse("recetas:presentacion_create", args=[receta.id])}
        )
        if active_presentaciones and presentacion_derivados_activos < len(active_presentaciones):
            chain_actions.append(
                {"label": "Sincronizar derivados", "url": reverse("recetas:receta_sync_derivados", args=[receta.id])}
            )
        if supply_chain_snapshot and not supply_chain_snapshot["has_downstream_usage"]:
            chain_actions.append(
                {
                    "label": "Crear producto final",
                    "url": f"{reverse('recetas:receta_create')}?mode=FINAL&source_base={receta.id}",
                }
            )
        sync_missing_count = max(len(active_presentaciones) - presentacion_derivados_activos, 0)
        base_chain_actions.append(
            {"label": "Administrar presentaciones", "url": reverse("recetas:presentacion_create", args=[receta.id])}
        )
        if not prep_insumo_operativo or sync_missing_count > 0:
            base_chain_actions.append(
                {
                    "label": "Sincronizar derivados",
                    "url": f"{reverse('recetas:receta_sync_derivados', args=[receta.id])}?next={reverse('recetas:receta_detail', args=[receta.id])}",
                }
            )
        if supply_chain_snapshot and not supply_chain_snapshot["has_downstream_usage"] and active_presentaciones:
            base_chain_actions.append(
                {
                    "label": "Crear producto final",
                    "url": f"{reverse('recetas:receta_create')}?mode=FINAL&source_base={receta.id}",
                }
            )
        base_chain_checkpoints = [
            {
                "label": "Insumo base derivado",
                "ok": prep_insumo_operativo,
                "detail": "Ya existe el insumo base maestro." if prep_insumo_operativo else "Falta generar o reactivar el derivado base.",
            },
            {
                "label": "Presentaciones activas",
                "ok": len(active_presentaciones) > 0,
                "detail": f"{len(active_presentaciones)} activa(s)." if active_presentaciones else "Todavía no hay tamaños activos.",
            },
            {
                "label": "Derivados activos",
                "ok": sync_missing_count == 0 and len(active_presentaciones) > 0,
                "detail": (
                    f"{presentacion_derivados_activos}/{len(active_presentaciones)} derivados listos."
                    if active_presentaciones
                    else "No aplica hasta crear presentaciones."
                ),
            },
            {
                "label": "Consumo final",
                "ok": bool(supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"]),
                "detail": (
                    f"{supply_chain_snapshot['downstream_recipe_count']} producto(s) final(es) usan esta base."
                    if supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"]
                    else "Aún no hay producto final consumiendo esta base."
                ),
            },
        ]
    else:
        next_steps = [
            "Agrega solo componentes maestros: panes, rellenos, coberturas, empaques o MP puntual.",
            "Usa copiar estructura si este producto se parece a otro ya armado.",
            "Ajusta cantidades finales; el sistema toma unidad y costo desde el maestro de artículos.",
        ]
    recetas_copiables = (
        Receta.objects.exclude(pk=receta.pk)
        .order_by("tipo", "nombre")
        .only("id", "nombre", "tipo", "familia", "categoria")
    )
    ordered_group_keys = (
        ["internos", "materia_prima", "empaques", "subsecciones", "pendientes"]
        if is_producto_final
        else ["materia_prima", "internos", "empaques", "subsecciones", "pendientes"]
    )
    group_meta = {
        "internos": {
            "title": "Insumos internos / subinsumos",
            "description": "Componentes producidos dentro de la empresa y reutilizados en esta receta.",
            "badge": "Interno",
        },
        "materia_prima": {
            "title": "Materia prima directa",
            "description": "Componentes comprados a proveedor y consumidos directamente en la receta.",
            "badge": "Materia prima",
        },
        "empaques": {
            "title": "Empaques",
            "description": "Material de presentación o empaque final ligado al costo del producto.",
            "badge": "Empaque",
        },
        "subsecciones": {
            "title": "Subsecciones / bloques",
            "description": "Bloques operativos de armado para ordenar coberturas, rellenos y decoración.",
            "badge": "Subsección",
        },
        "pendientes": {
            "title": "Sin artículo estándar",
            "description": "Componentes todavía sin artículo estándar asignado. Deben validarse para costeo correcto.",
            "badge": "Por validar",
        },
    }
    grouped_line_map: dict[str, list[LineaReceta]] = {key: [] for key in ordered_group_keys}
    direct_base_replacement_cache: dict[int, list[dict[str, object]]] = {}
    source_recipe_ids = {
        recipe_id
        for linea in lineas
        if linea.insumo_id
        for recipe_id in [_recipe_id_from_derived_code(linea.insumo.codigo or "")]
        if recipe_id
    }
    source_recipe_map = {
        item.id: item
        for item in Receta.objects.filter(id__in=source_recipe_ids).only("id", "nombre", "usa_presentaciones")
    }
    source_recipe_presentaciones = {
        row["receta_id"]: row["total"]
        for row in (
            RecetaPresentacion.objects.filter(receta_id__in=source_recipe_ids, activo=True)
            .values("receta_id")
            .annotate(total=Count("id"))
        )
    }
    for linea in lineas:
        _attach_linea_suggested_match(linea)
        _attach_linea_canonical_target(linea)
        linea.source_recipe = None
        linea.source_code_kind = None
        linea.source_active_presentaciones_count = 0
        linea.uses_direct_base_in_final = False
        linea.direct_base_replacement = None
        linea.erp_profile = (
            _insumo_operational_readiness(
                linea.insumo,
                ignore_supplier=is_producto_final,
            )
            if linea.insumo_id
            else None
        )
        if linea.insumo_id:
            linea.source_code_kind = _derived_code_kind(linea.insumo.codigo or "")
            source_recipe_id = _recipe_id_from_derived_code(linea.insumo.codigo or "")
            if source_recipe_id:
                linea.source_recipe = source_recipe_map.get(source_recipe_id)
                linea.source_active_presentaciones_count = int(
                    source_recipe_presentaciones.get(source_recipe_id, 0)
                )
                direct_base_candidate = bool(
                    is_producto_final
                    and linea.insumo.tipo_item == Insumo.TIPO_INTERNO
                    and linea.source_code_kind == "PREPARACION"
                    and linea.source_recipe
                    and linea.source_recipe.usa_presentaciones
                    and linea.source_active_presentaciones_count > 0
                )
                if direct_base_candidate:
                    direct_base_candidates = _active_presentation_derived_candidates(
                        source_recipe_id,
                        cache=direct_base_replacement_cache,
                    )
                    linea.direct_base_replacement = _suggest_direct_base_replacement(
                        linea,
                        cache=direct_base_replacement_cache,
                    )
                    linea.uses_direct_base_in_final = bool(linea.direct_base_replacement) or not direct_base_candidates
        grouped_line_map[_line_origin_bucket(linea)].append(linea)
    line_groups = [
        {
            "key": key,
            "title": group_meta[key]["title"],
            "description": group_meta[key]["description"],
            "badge": group_meta[key]["badge"],
            "lineas": grouped_line_map[key],
            "count": len(grouped_line_map[key]),
            "subtotal": sum((l.costo_total_estimado or 0.0) for l in grouped_line_map[key]),
            "erp_ready_count": sum(
                1
                for l in grouped_line_map[key]
                if getattr(l, "erp_profile", None) and l.erp_profile["ready"]
            ),
            "erp_incomplete_count": sum(
                1
                for l in grouped_line_map[key]
                if getattr(l, "erp_profile", None) and not l.erp_profile["ready"]
            ),
        }
        for key in ordered_group_keys
        if grouped_line_map[key]
    ]
    recipe_master_blockers = _recipe_master_blockers(lineas, receta=receta)
    recipe_master_gap_totals = _recipe_master_gap_totals_from_blockers(recipe_master_blockers)
    component_breakdown = []
    product_upstream_snapshot = _product_upstream_snapshot(lineas, receta=receta) if is_producto_final else None
    if product_upstream_snapshot is not None:
        setattr(receta, "_product_upstream_snapshot_cache", product_upstream_snapshot)
    derived_parent_snapshot = (
        product_upstream_snapshot.get("derived_parent_snapshot")
        if product_upstream_snapshot is not None
        else None
    )
    non_canonical_count = sum(1 for l in lineas if getattr(l, "canonical_needs_repoint", False))
    if total_costo_estimado > 0:
        if derived_parent_snapshot:
            derived_parent_cost = float(derived_parent_snapshot.get("parent_unit_cost") or 0.0)
            if derived_parent_cost > 0:
                component_breakdown.append(
                    {
                        "key": "parent_base_prorrated",
                        "title": "Base padre prorrateada",
                        "accent": "is-primary",
                        "subtotal": derived_parent_cost,
                        "percentage": (derived_parent_cost / total_costo_estimado * 100.0),
                        "count": 1,
                        "erp_ready_count": 1,
                        "erp_incomplete_count": 0,
                    }
                )
        breakdown_meta = {
            "internos": {"title": "Insumos internos", "accent": "is-success"},
            "materia_prima": {"title": "Materia prima puntual", "accent": "is-warning"},
            "empaques": {"title": "Empaques", "accent": "is-info"},
        }
        for key in ("internos", "materia_prima", "empaques"):
            subtotal = sum((l.costo_total_estimado or 0.0) for l in grouped_line_map[key])
            if subtotal <= 0:
                continue
            percentage = (subtotal / total_costo_estimado * 100.0) if total_costo_estimado else 0.0
            component_breakdown.append(
                {
                    "key": key,
                    "title": breakdown_meta[key]["title"],
                    "accent": breakdown_meta[key]["accent"],
                    "subtotal": subtotal,
                    "percentage": percentage,
                    "count": len(grouped_line_map[key]),
                    "erp_ready_count": sum(
                        1
                        for l in grouped_line_map[key]
                        if getattr(l, "erp_profile", None) and l.erp_profile["ready"]
                    ),
                    "erp_incomplete_count": sum(
                        1
                        for l in grouped_line_map[key]
                        if getattr(l, "erp_profile", None) and not l.erp_profile["ready"]
                    ),
                }
            )
    bom_integrity_alerts = []
    direct_base_usage_lines = [linea for linea in lineas if getattr(linea, "uses_direct_base_in_final", False)]
    direct_base_suggested_count = sum(
        1 for linea in direct_base_usage_lines if getattr(linea, "direct_base_replacement", None)
    )
    incomplete_item_count = sum(
        1 for l in lineas if getattr(l, "erp_profile", None) and not l.erp_profile["ready"]
    )
    prep_code = f"DERIVADO:RECETA:{receta.id}:PREPARACION"
    presentation_code_prefix = f"DERIVADO:RECETA:{receta.id}:PRESENTACION:"
    active_presentaciones = [p for p in presentaciones if p.activo]
    prep_insumo_ready = Insumo.objects.filter(codigo=prep_code, activo=True).exists()
    presentacion_derivados_activos = Insumo.objects.filter(
        codigo__startswith=presentation_code_prefix,
        activo=True,
    ).count()
    presentacion_health = None
    if receta.tipo == Receta.TIPO_PREPARACION:
        has_rendimiento = bool(receta.rendimiento_cantidad and receta.rendimiento_cantidad > 0)
        has_rendimiento_unidad = bool(receta.rendimiento_unidad_id)
        readiness_label = "Listo para derivados"
        readiness_level = "success"
        readiness_note = "La base ya tiene rendimiento, unidad y derivados suficientes para alimentar productos finales."
        if not has_rendimiento or not has_rendimiento_unidad:
            readiness_label = "Falta rendimiento base"
            readiness_level = "danger"
            readiness_note = "Captura rendimiento y unidad para calcular costo unitario y habilitar derivados consistentes."
        elif receta.usa_presentaciones and not active_presentaciones:
            readiness_label = "Sin presentaciones activas"
            readiness_level = "warning"
            readiness_note = "La receta está marcada con derivados, pero todavía no hay tamaños activos listos para usarse."
        elif receta.usa_presentaciones and presentacion_derivados_activos < len(active_presentaciones):
            readiness_label = "Derivados por sincronizar"
            readiness_level = "warning"
            readiness_note = "Hay presentaciones activas sin todos sus insumos derivados activos. Conviene revisar la sincronización."

        presentacion_health = {
            "readiness_label": readiness_label,
            "readiness_level": readiness_level,
            "readiness_note": readiness_note,
            "has_rendimiento": has_rendimiento,
            "has_rendimiento_unidad": has_rendimiento_unidad,
            "prep_insumo_ready": prep_insumo_ready,
            "active_presentaciones_count": len(active_presentaciones),
            "derived_presentaciones_count": presentacion_derivados_activos,
            "sync_recommended": bool(
                receta.usa_presentaciones
                and active_presentaciones
                and (not prep_insumo_ready or presentacion_derivados_activos < len(active_presentaciones))
            ),
        }
    if is_producto_final:
        if total_internos == 0 and not _recipe_has_effective_bom(receta):
            bom_integrity_alerts.append(
                {
                    "level": "danger",
                    "title": "Faltan insumos internos",
                    "description": "El producto final no tiene panes, rellenos, coberturas o subinsumos ligados. Esto suele indicar un BOM incompleto.",
                    "action_label": "Agregar insumo interno",
                    "action_url": f"{reverse('recetas:linea_create', args=[receta.id])}?component_kind=INSUMO_INTERNO&component_context=internos",
                }
            )
        if _recipe_requires_fixed_packaging(receta) and total_empaques == 0:
            bom_integrity_alerts.append(
                {
                    "level": "warning",
                    "title": "Sin empaque ligado",
                    "description": "No hay domos, cajas, etiquetas ni material final en el costeo. Revisa si el producto debe incluir empaque.",
                    "action_label": "Agregar empaque",
                    "action_url": f"{reverse('recetas:linea_create', args=[receta.id])}?component_kind=EMPAQUE&component_context=empaques",
                }
            )
        if total_lineas > 0 and total_internos == 0 and total_materia_prima > 0:
            bom_integrity_alerts.append(
                {
                    "level": "warning",
                    "title": "Producto armado directo con materia prima",
                    "description": "Este producto final está consumiendo solo materia prima puntual. Confirma que no deba usar insumos internos ya estandarizados.",
                }
            )
        if product_upstream_snapshot and product_upstream_snapshot["internal_without_source_count"] > 0:
            bom_integrity_alerts.append(
                {
                    "level": "warning",
                    "title": "Insumos internos sin base origen",
                    "description": f"Se detectaron {product_upstream_snapshot['internal_without_source_count']} línea(s) internas sin trazabilidad a receta base. Conviene normalizarlas para estabilizar costeo, MRP y compras.",
                }
            )
        if direct_base_usage_lines:
            ejemplos = ", ".join(
                dict.fromkeys(
                    (
                        getattr(linea.source_recipe, "nombre", "")
                        or getattr(linea.insumo, "nombre", "")
                    )
                    for linea in direct_base_usage_lines[:3]
                )
            )
            description = (
                f"Se detectaron {len(direct_base_usage_lines)} componente(s) usando la base completa "
                "aunque la receta origen ya maneja presentaciones activas. Sustituye esa base por la "
                "presentación derivada correcta para estabilizar costo y MRP."
            )
            if ejemplos:
                description += f" Bases afectadas: {ejemplos}."
            bom_integrity_alerts.append(
                {
                    "level": "danger",
                    "title": "Producto final usando base sin presentación",
                    "description": description,
                    "action_label": "Revisar componentes",
                    "action_url": reverse("recetas:receta_detail", args=[receta.id]),
                }
            )
        if total_revision or total_sin_match:
            bom_integrity_alerts.append(
                {
                    "level": "danger",
                    "title": "Hay componentes sin artículo estándar",
                    "description": "Mientras existan componentes por validar o sin artículo estándar, el costo y la planeación pueden quedar inestables.",
                    "action_label": "Resolver catálogo",
                    "action_url": f"{reverse('recetas:matching_pendientes')}?receta={receta.id}",
                }
            )
        if non_canonical_count:
            bom_integrity_alerts.append(
                {
                    "level": "warning",
                    "title": "Hay artículos fuera de maestro",
                    "description": f"Se detectaron {non_canonical_count} línea(s) apuntando a variantes fuera de maestro. Conviene corregirlas para estabilizar costos, compras y reportes.",
                    "action_label": "Usar artículo estándar",
                    "action_url": reverse("recetas:receta_detail", args=[receta.id]),
                }
            )
        if incomplete_item_count:
            bom_integrity_alerts.append(
                {
                    "level": "danger" if recipe_master_gap_totals["critical"] > 0 else "warning",
                    "title": (
                        "Demanda crítica bloqueada por maestro"
                        if recipe_master_gap_totals["critical"] > 0
                        else "Hay artículos incompletos en maestro"
                    ),
                    "description": (
                        f"Se detectaron {recipe_master_gap_totals['critical']} artículo(s) críticos por demanda con brechas en maestro. Conviene corregirlos antes de cerrar costeo, MRP o compras."
                        if recipe_master_gap_totals["critical"] > 0
                        else f"Se detectaron {incomplete_item_count} componente(s) ligados a artículos todavía incompletos en el maestro. Conviene corregirlos antes de cerrar costeo o MRP."
                    ),
                    "action_label": "Revisar maestro",
                    "action_url": f"{reverse('maestros:insumo_list')}?enterprise_status=incompletos&usage_scope=recipes&recipe_scope=finales",
                }
            )
    elif is_base_con_presentaciones:
        if not receta.rendimiento_cantidad:
            bom_integrity_alerts.append(
                {
                    "level": "danger",
                    "title": "Falta rendimiento base",
                    "description": "Sin rendimiento no puede calcularse el costo por unidad ni el costo de las presentaciones derivadas.",
                    "action_label": "Abrir receta",
                    "action_url": reverse("recetas:receta_detail", args=[receta.id]),
                }
            )
        if not presentaciones:
            bom_integrity_alerts.append(
                {
                    "level": "warning",
                    "title": "Sin presentaciones activas",
                    "description": "La receta está marcada para derivados, pero aún no tiene presentaciones configuradas.",
                    "action_label": "Agregar presentación",
                    "action_url": reverse("recetas:presentacion_create", args=[receta.id]),
                }
            )
        elif presentacion_derivados_activos < len(active_presentaciones):
            bom_integrity_alerts.append(
                {
                    "level": "warning",
                    "title": "Derivados pendientes de sincronizar",
                    "description": "Hay presentaciones activas que todavía no generaron todos sus insumos derivados maestros. Conviene sincronizar antes de usarlas en productos finales.",
                    "action_label": "Sincronizar derivados",
                    "action_url": reverse("recetas:receta_sync_derivados", args=[receta.id]),
                }
            )
        elif not _recipe_supply_chain_snapshot(receta)["has_downstream_usage"]:
            bom_integrity_alerts.append(
                {
                    "level": "warning",
                    "title": "Sin consumo final",
                    "description": "La base ya tiene derivados, pero todavía no se consume en ningún producto final. Conviene cerrar la cadena operativa.",
                    "action_label": "Crear producto final",
                    "action_url": f"{reverse('recetas:receta_create')}?mode=FINAL&source_base={receta.id}",
                }
            )
    elif is_base_simple and not receta.rendimiento_cantidad:
        bom_integrity_alerts.append(
            {
                "level": "warning",
                "title": "Falta rendimiento",
                "description": "Captura el rendimiento total para que el costo unitario de la batida quede bien calculado.",
                "action_label": "Abrir receta",
                "action_url": reverse("recetas:receta_detail", args=[receta.id]),
            }
        )
    recipe_chain_status = _recipe_chain_focus_summary(receta)
    governance_issues = _recipe_governance_issues(receta)
    release_gate_rows = [
        {
            "label": "Maestro ERP cerrado",
            "status": recipe_master_gap_totals["total"] == 0,
            "detail": (
                "Todos los artículos ligados ya tienen la base maestra completa."
                if recipe_master_gap_totals["total"] == 0
                else f'{recipe_master_gap_totals["total"]} artículo(s) aún requieren datos maestros.'
            ),
            "action_label": "Abrir maestro",
            "action_url": reverse("maestros:insumo_list"),
        },
        {
            "label": "BOM sin componentes abiertos",
            "status": total_revision == 0 and total_sin_match == 0,
            "detail": (
                "La estructura ya quedó integrada y sin componentes abiertos."
                if total_revision == 0 and total_sin_match == 0
                else f"{total_revision} componente(s) por validar · {total_sin_match} sin artículo estándar."
            ),
            "action_label": add_line_label,
            "action_url": reverse("recetas:linea_create", args=[receta.id]),
        },
        {
            "label": "Costo operativo calculado",
            "status": total_costo_estimado > 0,
            "detail": (
                f"Costo operativo actual ${total_costo_estimado:.2f}."
                if total_costo_estimado > 0
                else "Todavía no hay costo operativo suficiente para cerrar el documento."
            ),
            "action_label": "Abrir costeo",
            "action_url": reverse("recetas:drivers_costeo"),
        },
        {
            "label": "Cadena ERP cerrada",
            "status": recipe_chain_status["tone"] != "danger",
            "detail": recipe_chain_status["summary"],
            "action_label": chain_actions[0]["label"] if chain_actions else "Revisar documento",
            "action_url": chain_actions[0]["url"] if chain_actions else reverse("recetas:receta_detail", args=[receta.id]),
        },
    ]
    release_gate_progress = {
        "completed": sum(1 for row in release_gate_rows if row["status"]),
        "total": len(release_gate_rows),
    }
    erp_governance_rows = _recipe_detail_governance_rows(
        receta=receta,
        recipe_master_gap_totals=recipe_master_gap_totals,
        total_revision=total_revision,
        total_sin_match=total_sin_match,
        release_gate_progress=release_gate_progress,
        chain_focus_summary=chain_focus_summary,
    )
    trunk_handoff_rows = _recipe_detail_trunk_handoff_rows(
        receta=receta,
        recipe_master_gap_totals=recipe_master_gap_totals,
        total_revision=total_revision,
        total_sin_match=total_sin_match,
        release_gate_progress=release_gate_progress,
        chain_focus_summary=chain_focus_summary,
        is_producto_final=is_producto_final,
        is_base_con_presentaciones=is_base_con_presentaciones,
        governance_issues=governance_issues,
        supply_chain_snapshot=supply_chain_snapshot,
        product_upstream_snapshot=product_upstream_snapshot,
    )
    erp_control_chain = [
        {
            "label": "01 Maestro",
            "status": "Listo" if recipe_master_gap_totals["total"] == 0 else "Con brechas",
            "tone": "success" if recipe_master_gap_totals["total"] == 0 else "warning",
            "detail": (
                "Los artículos ligados ya cumplen datos base."
                if recipe_master_gap_totals["total"] == 0
                else f'{recipe_master_gap_totals["total"]} artículo(s) requieren cierre maestro.'
            ),
            "action_label": "Abrir maestro",
            "action_url": reverse("maestros:insumo_list"),
        },
        {
            "label": "02 BOM",
            "status": "Listo" if total_revision == 0 and total_sin_match == 0 else "Por validar",
            "tone": "success" if total_revision == 0 and total_sin_match == 0 else "warning",
            "detail": (
                f"{total_lineas} componente(s) ya estructurados."
                if total_revision == 0 and total_sin_match == 0
                else f"{total_revision} componente(s) por validar · {total_sin_match} sin artículo estándar."
            ),
            "action_label": add_line_label,
            "action_url": reverse("recetas:linea_create", args=[receta.id]),
        },
        {
            "label": "03 Costeo",
            "status": "Calculado" if total_costo_estimado > 0 else "Por validar",
            "tone": "success" if total_costo_estimado > 0 else "warning",
            "detail": (
                f"Costo operativo actual ${total_costo_estimado:.2f}."
                if total_costo_estimado > 0
                else "Todavía no hay costo operativo suficiente para cerrar el documento."
            ),
            "action_label": "Abrir costeo",
            "action_url": reverse("recetas:drivers_costeo"),
        },
        {
            "label": "04 Liberación",
            "status": (
                "Lista"
                if recipe_master_gap_totals["total"] == 0
                and total_revision == 0
                and total_sin_match == 0
                and total_costo_estimado > 0
                and recipe_chain_status["tone"] != "danger"
                else "Por validar"
            ),
            "tone": (
                "success"
                if recipe_master_gap_totals["total"] == 0
                and total_revision == 0
                and total_sin_match == 0
                and total_costo_estimado > 0
                and recipe_chain_status["tone"] != "danger"
                else "warning"
            ),
            "detail": (
                "Documento listo para costeo, planeación y operación diaria."
                if recipe_master_gap_totals["total"] == 0
                and total_revision == 0
                and total_sin_match == 0
                and total_costo_estimado > 0
                and recipe_chain_status["tone"] != "danger"
                else recipe_chain_status["summary"]
            ),
            "action_label": chain_actions[0]["label"] if chain_actions else "Revisar documento",
            "action_url": chain_actions[0]["url"] if chain_actions else reverse("recetas:receta_detail", args=[receta.id]),
        },
    ]
    module_enablement_cards: list[dict[str, str | bool]] = []
    master_ready = recipe_master_gap_totals["total"] == 0
    bom_ready = total_revision == 0 and total_sin_match == 0
    cost_ready = total_costo_estimado > 0
    chain_ready = recipe_chain_status["tone"] != "danger"
    packaging_ready = total_empaques > 0 or (is_producto_final and not _recipe_requires_fixed_packaging(receta))
    internal_ready = total_internos > 0
    if is_producto_final:
        module_enablement_cards = [
            {
                "label": "Costeo",
                "status": master_ready and bom_ready and cost_ready,
                "tone": "success" if master_ready and bom_ready and cost_ready else "warning",
                "detail": (
                    "El documento ya tiene base maestra, BOM estable y costo repetible."
                    if master_ready and bom_ready and cost_ready
                    else "Todavía falta cerrar maestro, estructura o costo para usar este producto como referencia financiera."
                ),
                "action_label": "Abrir costeo" if not cost_ready else "Revisar documento",
                "action_url": reverse("recetas:drivers_costeo") if not cost_ready else reverse("recetas:receta_detail", args=[receta.id]),
            },
            {
                "label": "MRP",
                "status": master_ready and bom_ready and cost_ready and chain_ready and internal_ready,
                "tone": "success" if master_ready and bom_ready and cost_ready and chain_ready and internal_ready else "warning",
                "detail": (
                    "El producto ya puede explotar demanda y reabasto con trazabilidad cerrada."
                    if master_ready and bom_ready and cost_ready and chain_ready and internal_ready
                    else "MRP sigue bloqueado hasta cerrar componentes internos, maestro y trazabilidad del documento."
                ),
                "action_label": "Abrir MRP" if master_ready and bom_ready and cost_ready and chain_ready and internal_ready else "Cerrar documento",
                "action_url": reverse("recetas:mrp_form") if master_ready and bom_ready and cost_ready and chain_ready and internal_ready else reverse("recetas:receta_detail", args=[receta.id]),
            },
            {
                "label": "Compras",
                "status": master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready,
                "tone": "success" if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready else "warning",
                "detail": (
                    "La receta ya puede alimentar solicitudes y órdenes sin ambigüedad documental."
                    if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready
                    else "Compras sigue bloqueado hasta cerrar empaque, estructura BOM y artículos del maestro."
                ),
                "action_label": "Abrir compras" if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready else "Revisar bloqueos",
                "action_url": reverse("compras:solicitudes") if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready else reverse("recetas:receta_detail", args=[receta.id]),
            },
            {
                "label": "Inventario",
                "status": master_ready and chain_ready and packaging_ready,
                "tone": "success" if master_ready and chain_ready and packaging_ready else "warning",
                "detail": (
                    "Inventario ya puede trabajar con el producto y su empaque como artículo operativo estable."
                    if master_ready and chain_ready and packaging_ready
                    else "Inventario aún requiere maestro completo, trazabilidad cerrada o empaque definido."
                ),
                "action_label": "Abrir inventario" if master_ready and chain_ready and packaging_ready else "Revisar maestro",
                "action_url": reverse("inventario:existencias") if master_ready and chain_ready and packaging_ready else reverse("maestros:insumo_list"),
            },
            {
                "label": "Venta final",
                "status": master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready,
                "tone": "success" if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready else "warning",
                "detail": (
                    "El artículo ya está en condición de venta final consistente."
                    if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready
                    else "Todavía no conviene liberar este artículo a operación final hasta cerrar estructura, costo y empaque."
                ),
                "action_label": "Revisar catálogo" if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready else "Cerrar ERP",
                "action_url": reverse("recetas:recetas_list",) if master_ready and bom_ready and cost_ready and chain_ready and internal_ready and packaging_ready else reverse("recetas:receta_detail", args=[receta.id]),
            },
        ]
    elif is_base_con_presentaciones:
        module_enablement_cards = [
            {
                "label": "Costeo base",
                "status": bool(presentacion_health and presentacion_health["has_rendimiento"] and presentacion_health["has_rendimiento_unidad"]),
                "tone": "success" if presentacion_health and presentacion_health["has_rendimiento"] and presentacion_health["has_rendimiento_unidad"] else "warning",
                "detail": (
                    "La base ya tiene rendimiento y unidad para costeo repetible."
                    if presentacion_health and presentacion_health["has_rendimiento"] and presentacion_health["has_rendimiento_unidad"]
                    else "Todavía falta rendimiento o unidad para que la base costee correctamente."
                ),
                "action_label": "Abrir base",
                "action_url": reverse("recetas:receta_detail", args=[receta.id]),
            },
            {
                "label": "Derivados",
                "status": bool(presentacion_health and presentacion_health["active_presentaciones_count"] > 0 and presentacion_health["derived_presentaciones_count"] >= presentacion_health["active_presentaciones_count"]),
                "tone": "success" if presentacion_health and presentacion_health["active_presentaciones_count"] > 0 and presentacion_health["derived_presentaciones_count"] >= presentacion_health["active_presentaciones_count"] else "warning",
                "detail": (
                    "Las presentaciones activas ya quedaron sincronizadas como artículos internos."
                    if presentacion_health and presentacion_health["active_presentaciones_count"] > 0 and presentacion_health["derived_presentaciones_count"] >= presentacion_health["active_presentaciones_count"]
                    else "Todavía faltan presentaciones activas o sincronización de derivados."
                ),
                "action_label": "Administrar presentaciones",
                "action_url": reverse("recetas:presentacion_create", args=[receta.id]),
            },
            {
                "label": "Producto final",
                "status": bool(supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"]),
                "tone": "success" if supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"] else "warning",
                "detail": (
                    "Esta base ya alimenta productos finales y está integrada en la cadena comercial."
                    if supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"]
                    else "La base aún no alimenta productos finales; la cadena sigue abierta."
                ),
                "action_label": "Crear final" if not (supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"]) else "Revisar cadena",
                "action_url": f"{reverse('recetas:receta_create')}?mode=FINAL&source_base={receta.id}" if not (supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"]) else reverse("recetas:receta_detail", args=[receta.id]),
            },
            {
                "label": "MRP / Compras",
                "status": bool(supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"] and presentacion_health and presentacion_health["active_presentaciones_count"] > 0),
                "tone": "success" if supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"] and presentacion_health and presentacion_health["active_presentaciones_count"] > 0 else "warning",
                "detail": (
                    "La base ya está en condición de alimentar planeación y abastecimiento aguas abajo."
                    if supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"] and presentacion_health and presentacion_health["active_presentaciones_count"] > 0
                    else "MRP y compras solo deben usar esta base cuando derivados y consumo final queden cerrados."
                ),
                "action_label": "Abrir MRP" if supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"] else "Cerrar cadena",
                "action_url": reverse("recetas:mrp_form") if supply_chain_snapshot and supply_chain_snapshot["has_downstream_usage"] else reverse("recetas:receta_detail", args=[receta.id]),
            },
        ]
    else:
        module_enablement_cards = [
            {
                "label": "Costeo base",
                "status": bool(receta.rendimiento_cantidad and receta.rendimiento_unidad_id and cost_ready),
                "tone": "success" if receta.rendimiento_cantidad and receta.rendimiento_unidad_id and cost_ready else "warning",
                "detail": (
                    "La batida base ya tiene costo unitario estable."
                    if receta.rendimiento_cantidad and receta.rendimiento_unidad_id and cost_ready
                    else "Todavía falta rendimiento, unidad o costo para estabilizar la base."
                ),
                "action_label": "Abrir base",
                "action_url": reverse("recetas:receta_detail", args=[receta.id]),
            },
            {
                "label": "Derivados opcionales",
                "status": not receta.usa_presentaciones,
                "tone": "success" if not receta.usa_presentaciones else "warning",
                "detail": (
                    "Esta base opera sin tamaños derivados, por lo que ya puede seguir como insumo interno simple."
                    if not receta.usa_presentaciones
                    else "La base quedó marcada con presentaciones; conviene cerrarlas antes de usarla aguas abajo."
                ),
                "action_label": "Revisar base",
                "action_url": reverse("recetas:receta_detail", args=[receta.id]),
            },
            {
                "label": "Uso operativo",
                "status": master_ready and bom_ready and cost_ready,
                "tone": "success" if master_ready and bom_ready and cost_ready else "warning",
                "detail": (
                    "La base ya puede operar como insumo interno consistente."
                    if master_ready and bom_ready and cost_ready
                    else "Todavía falta cerrar maestro, estructura o costo para usar esta base con seguridad."
                ),
                "action_label": "Abrir documento",
                "action_url": reverse("recetas:receta_detail", args=[receta.id]),
            },
        ]
    module_handoff_rows = _module_enablement_handoff_rows(receta, module_enablement_cards)
    return render(
        request,
        "recetas/receta_detail.html",
        {
            "receta": receta,
            "lineas": lineas,
            "presentaciones": presentaciones,
            "total_lineas": total_lineas,
            "total_match": total_match,
            "total_revision": total_revision,
            "total_sin_match": total_sin_match,
            "total_costo_estimado": total_costo_estimado,
            "total_materia_prima": total_materia_prima,
            "total_internos": total_internos,
            "total_empaques": total_empaques,
            "total_subsecciones": total_subsecciones,
            "receta_rol_label": receta_rol_label,
            "receta_rol_description": receta_rol_description,
            "is_producto_final": is_producto_final,
            "is_base_con_presentaciones": is_base_con_presentaciones,
            "is_base_simple": is_base_simple,
            "add_line_label": add_line_label,
            "copy_label": copy_label,
            "next_steps": next_steps,
            "chain_actions": chain_actions,
            "chain_focus_summary": chain_focus_summary,
            "enterprise_stage": enterprise_stage,
            "enterprise_stage_playbook": enterprise_stage_playbook,
            "unidades": UnidadMedida.objects.order_by("codigo"),
            "tipo_choices": Receta.TIPO_CHOICES,
            "costo_por_kg_estimado": receta.costo_por_kg_estimado,
            "linea_tipo_choices": LineaReceta.TIPO_CHOICES,
            "costeo_actual": costeo_actual,
            "costeo_unavailable": costeo_unavailable,
            "versiones_recientes": versiones_recientes,
            "versiones_all": versiones_all,
            "versiones_comparativo": comparativo,
            "versiones_unavailable": versiones_unavailable,
            "selected_base": selected_base,
            "selected_target": selected_target,
            "version_compare": compare_data,
            "rend_unit_code": (receta.rendimiento_unidad.codigo if receta.rendimiento_unidad else ""),
            "familias_catalogo": familias_catalogo,
            "categorias_catalogo": categorias_catalogo,
            "familia_categoria_catalogo_json": familia_categoria_catalogo_json(categorias_catalogo),
            "recetas_copiables": recetas_copiables,
            "line_groups": line_groups,
            "recipe_master_blockers": recipe_master_blockers,
            "recipe_master_gap_totals": recipe_master_gap_totals,
            "component_breakdown": component_breakdown,
            "product_upstream_snapshot": product_upstream_snapshot,
            "derived_parent_snapshot": derived_parent_snapshot,
            "total_costo_directo": total_costo_directo,
            "bom_integrity_alerts": bom_integrity_alerts,
            "non_canonical_count": non_canonical_count,
            "direct_base_suggested_count": direct_base_suggested_count,
            "presentacion_health": presentacion_health,
            "supply_chain_snapshot": supply_chain_snapshot,
            "base_chain_actions": base_chain_actions,
            "base_chain_checkpoints": base_chain_checkpoints,
            "erp_control_chain": erp_control_chain,
            "trunk_handoff_rows": trunk_handoff_rows,
            "trunk_handoff_summary": _trunk_handoff_summary(
                trunk_handoff_rows,
                owner="Recetas / Costeo",
                fallback_url=reverse("recetas:receta_detail", args=[receta.id]),
            ),
            "module_enablement_cards": module_enablement_cards,
            "module_handoff_rows": module_handoff_rows,
            "release_gate_rows": release_gate_rows,
            "release_gate_progress": release_gate_progress,
            "erp_command_center": {
                "owner": "Recetas / Costeo",
                "status": "Crítico"
                if (total_revision + total_sin_match + recipe_master_gap_totals["total"]) > 0
                else "Seguimiento"
                if (
                    int(round((release_gate_progress["completed"] / release_gate_progress["total"]) * 100))
                    if release_gate_progress["total"]
                    else 0
                ) < 100
                else "Controlado",
                "tone": "danger"
                if (total_revision + total_sin_match + recipe_master_gap_totals["total"]) > 0
                else "warning"
                if (
                    int(round((release_gate_progress["completed"] / release_gate_progress["total"]) * 100))
                    if release_gate_progress["total"]
                    else 0
                ) < 100
                else "success",
                "blockers": total_revision + total_sin_match + recipe_master_gap_totals["total"],
                "next_step": chain_focus_summary.get("detail")
                or "Continuar el cierre del documento y su trazabilidad operativa.",
                "url": chain_focus_summary.get("action_url") or reverse("recetas:receta_detail", args=[receta.id]),
                "cta": chain_focus_summary.get("action_label") or "Abrir documento",
            },
            "erp_governance_rows": erp_governance_rows,
            "executive_radar_rows": _recipes_executive_radar_rows(
                erp_governance_rows,
                owner="Recetas / Costeo",
                fallback_url=reverse("recetas:receta_detail", args=[receta.id]),
            ),
            "critical_path_rows": _recipes_critical_path_rows(
                erp_governance_rows,
                owner="Recetas / Costeo",
                fallback_url=reverse("recetas:receta_detail", args=[receta.id]),
            ),
        },
    )


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def receta_update(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    nombre = (request.POST.get("nombre") or "").strip()
    codigo_point = (request.POST.get("codigo_point") or "").strip()
    familia = (request.POST.get("familia") or "").strip()
    categoria = (request.POST.get("categoria") or "").strip()
    sheet_name = (request.POST.get("sheet_name") or "").strip()
    tipo = (request.POST.get("tipo") or Receta.TIPO_PREPARACION).strip()
    usa_presentaciones = request.POST.get("usa_presentaciones") == "on"
    rendimiento_cantidad = _to_decimal_or_none(request.POST.get("rendimiento_cantidad"))
    rendimiento_unidad_id = request.POST.get("rendimiento_unidad_id")

    if not nombre:
        messages.error(request, "El nombre de receta es obligatorio.")
        return redirect("recetas:receta_detail", pk=pk)

    if tipo not in {Receta.TIPO_PREPARACION, Receta.TIPO_PRODUCTO_FINAL}:
        tipo = Receta.TIPO_PREPARACION

    validation_errors = _validate_receta_enterprise_rules(
        tipo=tipo,
        usa_presentaciones=usa_presentaciones,
        rendimiento_cantidad=rendimiento_cantidad,
        rendimiento_unidad_id=rendimiento_unidad_id,
        familia=familia,
        presentaciones_existentes=receta.presentaciones.count(),
    )
    if validation_errors:
        for error in validation_errors:
            messages.error(request, error)
        return redirect("recetas:receta_detail", pk=pk)

    if tipo == Receta.TIPO_PRODUCTO_FINAL:
        usa_presentaciones = False
        rendimiento_cantidad = None
        rendimiento_unidad_id = ""

    receta.nombre = nombre[:250]
    receta.codigo_point = codigo_point[:80]
    receta.familia = familia[:120]
    receta.categoria = categoria[:120]
    receta.sheet_name = sheet_name[:120]
    receta.tipo = tipo
    receta.usa_presentaciones = usa_presentaciones
    receta.rendimiento_cantidad = rendimiento_cantidad
    receta.rendimiento_unidad = UnidadMedida.objects.filter(pk=rendimiento_unidad_id).first() if rendimiento_unidad_id else None
    receta.save()
    _sync_derived_insumos_safe(request, receta)
    _sync_cost_version_safe(request, receta, "RECETA_UPDATE")
    messages.success(request, "Receta actualizada.")
    return redirect("recetas:receta_detail", pk=pk)


def _receta_delete_blockers(receta: Receta) -> list[tuple[str, int]]:
    blockers = [
        ("ventas históricas", receta.ventas_historicas.count()),
        ("ventas POS", receta.ventas_pos.count()),
        ("mermas POS", receta.mermas_pos.count()),
        ("solicitudes de venta", receta.solicitudes_venta.count()),
        ("pronósticos", receta.pronosticos.count()),
        ("renglones de plan", receta.plan_items.count()),
        ("políticas de stock por sucursal", receta.politicas_stock_sucursal.count()),
        ("movimientos CEDIS", receta.movimientos_cedis.count()),
        ("líneas de reabasto CEDIS", receta.solicitudes_reabasto_lineas.count()),
        ("reservas pickup", receta.pickup_reservations.count()),
        ("ventas diarias Point", receta.point_daily_sales.count()),
    ]
    if getattr(receta, "inventario_cedis", None):
        blockers.append(("inventario CEDIS", 1))
    return [(label, count) for label, count in blockers if count]


@login_required
@permission_required("recetas.delete_receta", raise_exception=True)
@require_POST
def receta_delete(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    blockers = _receta_delete_blockers(receta)
    if blockers:
        summary = ", ".join(f"{count} {label}" for label, count in blockers[:4])
        extra = f" y {len(blockers) - 4} más" if len(blockers) > 4 else ""
        messages.error(
            request,
            f"No se puede eliminar la receta porque ya tiene huella operativa: {summary}{extra}.",
        )
        return redirect("recetas:receta_detail", pk=pk)

    nombre = receta.nombre
    receta_id = receta.id
    try:
        receta.delete()
    except ProtectedError:
        messages.error(
            request,
            "No se puede eliminar la receta porque todavía está ligada a documentos operativos protegidos.",
        )
        return redirect("recetas:receta_detail", pk=pk)

    log_event(
        request.user,
        "DELETE",
        "recetas.Receta",
        str(receta_id),
        {"receta_id": receta_id, "nombre": nombre},
    )
    messages.success(request, f"Receta eliminada: {nombre}.")
    return redirect("recetas:recetas_list")


def _to_decimal_or_none(value: str | None) -> Decimal | None:
    if value is None:
        return None
    raw = str(value).strip().replace(",", ".")
    if raw == "":
        return None
    try:
        return Decimal(raw)
    except Exception:
        return None


def _to_non_negative_decimal_or_none(value: str | None) -> Decimal | None:
    if value is None:
        return None
    raw = str(value).strip().replace(",", ".")
    if raw == "":
        return None
    try:
        num = Decimal(raw)
        return num if num >= 0 else None
    except Exception:
        return None


def _validate_receta_enterprise_rules(
    *,
    mode_code: str | None = None,
    tipo: str | None = None,
    usa_presentaciones: bool = False,
    rendimiento_cantidad: Decimal | None = None,
    rendimiento_unidad_id: str | int | None = None,
    familia: str = "",
    presentaciones_existentes: int = 0,
) -> list[str]:
    errors: list[str] = []
    normalized_mode = (mode_code or "").strip().upper()
    receta_tipo = (tipo or Receta.TIPO_PREPARACION).strip()
    is_producto_final = normalized_mode == "FINAL" or receta_tipo == Receta.TIPO_PRODUCTO_FINAL
    is_base_con_presentaciones = normalized_mode == "BASE_DERIVADOS" or (
        receta_tipo == Receta.TIPO_PREPARACION and usa_presentaciones
    )
    is_base_simple = normalized_mode == "BASE" or (
        receta_tipo == Receta.TIPO_PREPARACION and not usa_presentaciones
    )

    has_rendimiento = bool(rendimiento_cantidad and rendimiento_cantidad > 0)
    has_rendimiento_unidad = bool(str(rendimiento_unidad_id or "").strip())

    if is_producto_final:
        if not familia:
            errors.append("En producto final debes seleccionar una familia.")
        if usa_presentaciones:
            errors.append("Un producto final no puede usar presentaciones derivadas.")
        if presentaciones_existentes:
            errors.append(
                "No puedes convertir a producto final una receta que aún tiene presentaciones activas. Elimínalas o consolídalas primero."
            )
        return errors

    if is_base_simple and presentaciones_existentes:
        errors.append(
            "No puedes convertir a base simple una receta que ya tiene presentaciones activas. Elimínalas primero o conserva el modo con derivados."
        )

    if is_base_simple or is_base_con_presentaciones:
        if not has_rendimiento:
            errors.append("Debes capturar el rendimiento total de la batida para costeo enterprise.")
        if not has_rendimiento_unidad:
            errors.append("Debes seleccionar la unidad del rendimiento para costeo enterprise.")

    return errors


def _validate_presentacion_workflow(receta: Receta) -> list[str]:
    errors: list[str] = []
    if receta.tipo != Receta.TIPO_PREPARACION:
        errors.append("Solo un insumo base permite presentaciones derivadas.")
        return errors
    if not receta.rendimiento_cantidad or receta.rendimiento_cantidad <= 0:
        errors.append("Antes de administrar presentaciones debes capturar el rendimiento total de la base.")
    if not receta.rendimiento_unidad_id:
        errors.append("Antes de administrar presentaciones debes capturar la unidad del rendimiento.")
    return errors


def _linea_form_context(
    receta: Receta,
    linea: LineaReceta | None = None,
    *,
    advanced_mode: bool = False,
    component_filter: str | None = None,
    component_context: str | None = None,
) -> Dict[str, Any]:
    raw_insumos = [
        insumo
        for insumo in canonicalized_insumo_selector(limit=1600)
        if insumo.activo and insumo.unidad_base_id
    ]
    selector_source_recipe_ids = {
        recipe_id
        for insumo in raw_insumos
        for recipe_id in [_recipe_id_from_derived_code(insumo.codigo or "")]
        if recipe_id
    }
    selector_source_recipe_map = {
        item.id: item
        for item in Receta.objects.filter(id__in=selector_source_recipe_ids).only("id", "usa_presentaciones")
    }
    selector_source_presentaciones = {
        row["receta_id"]: row["total"]
        for row in (
            RecetaPresentacion.objects.filter(receta_id__in=selector_source_recipe_ids, activo=True)
            .values("receta_id")
            .annotate(total=Count("id"))
        )
    }

    generic_recipe_tokens = {
        "pastel",
        "pasteles",
        "producto",
        "final",
        "base",
        "batida",
        "mezcla",
        "chico",
        "mediano",
        "grande",
        "mini",
        "bollos",
        "bollo",
        "individual",
        "rosca",
        "media",
        "plancha",
        "de",
        "del",
        "con",
        "para",
        "y",
    }
    recipe_context_tokens = [
        token
        for token in normalizar_nombre(f"{receta.nombre} {receta.familia or ''} {receta.categoria or ''}").split()
        if len(token) >= 3 and token not in generic_recipe_tokens
    ]

    def _recipe_context_score(i: Insumo) -> int:
        if not recipe_context_tokens:
            return 0
        haystack = normalizar_nombre(" ".join([i.nombre or "", i.categoria or "", i.codigo or "", i.codigo_point or ""]))
        score = 0
        for token in recipe_context_tokens:
            if token and token in haystack:
                score += 1
        return score

    def _option_score(i: Insumo) -> int:
        score = 0
        code = (i.codigo or "")
        derived_kind = _derived_code_kind(code)
        source_recipe_id = _recipe_id_from_derived_code(code)
        source_recipe = selector_source_recipe_map.get(source_recipe_id) if source_recipe_id else None
        source_active_presentaciones = int(selector_source_presentaciones.get(source_recipe_id, 0)) if source_recipe_id else 0
        is_internal = i.tipo_item == Insumo.TIPO_INTERNO or code.startswith("DERIVADO:RECETA:")
        is_empaque = i.tipo_item == Insumo.TIPO_EMPAQUE or "empaque" in normalizar_nombre(i.categoria or "")
        if is_internal:
            score += 300
            if derived_kind == "PREPARACION":
                score += 50
            elif derived_kind == "PRESENTACION":
                score += 120
        elif is_empaque:
            score += 120
        if (
            receta.tipo == Receta.TIPO_PRODUCTO_FINAL
            and derived_kind == "PREPARACION"
            and source_recipe
            and source_recipe.usa_presentaciones
            and source_active_presentaciones > 0
        ):
            score -= 240
        if (
            receta.tipo == Receta.TIPO_PRODUCTO_FINAL
            and derived_kind == "PRESENTACION"
            and source_recipe
            and source_recipe.usa_presentaciones
            and source_active_presentaciones > 0
        ):
            score += 180
        if (i.codigo_point or "").strip():
            score += 80
        if i.latest_costo_unitario is not None and Decimal(str(i.latest_costo_unitario or 0)) > 0:
            score += 60
        erp_profile = getattr(i, "erp_profile", None) or _insumo_erp_readiness(i)
        if erp_profile["ready"]:
            score += 140
        else:
            score -= 160
        score += _recipe_context_score(i) * 45
        if i.proveedor_principal_id:
            score += 20
        if i.unidad_base_id:
            score += 10
        # Evita que "ruido" de pruebas prevalezca sobre catálogo real.
        if normalizar_nombre(i.nombre).startswith("qa flow"):
            score -= 120
        return score

    for insumo in raw_insumos:
        if not hasattr(insumo, "origen_orden"):
            insumo.origen_orden = (
                0
                if insumo.tipo_item == Insumo.TIPO_INTERNO or (insumo.codigo or "").startswith("DERIVADO:RECETA:")
                else 1
                if insumo.tipo_item == Insumo.TIPO_EMPAQUE
                else 2
            )
        if not hasattr(insumo, "canonical_variant_count"):
            insumo.canonical_variant_count = len(getattr(insumo, "member_ids", []) or [insumo.id])
        insumo.recipe_context_score = _recipe_context_score(insumo)
        insumo.erp_profile = _insumo_erp_readiness(insumo)
        insumo.source_recipe_id = _recipe_id_from_derived_code(insumo.codigo or "")
        insumo.source_code_kind = _derived_code_kind(insumo.codigo or "")
        insumo.source_active_presentaciones = int(
            selector_source_presentaciones.get(insumo.source_recipe_id, 0)
        ) if insumo.source_recipe_id else 0
        insumo.option_score = _option_score(insumo)

    insumos = sorted(
        raw_insumos,
        key=lambda x: ((x.origen_orden or 9), -(x.option_score or 0), x.nombre.lower()),
    )
    if linea and linea.insumo_id:
        selected_present = any(i.id == linea.insumo_id for i in insumos)
        if not selected_present and linea.insumo and linea.insumo.activo and linea.insumo.unidad_base_id:
            linea.insumo.erp_profile = _insumo_erp_readiness(linea.insumo)
            linea.insumo.option_score = _option_score(linea.insumo)
            insumos.append(linea.insumo)
            insumos = sorted(
                insumos,
                key=lambda x: ((x.origen_orden or 9), -(getattr(x, "option_score", 0) or 0), x.nombre.lower()),
            )

    def _is_internal(i: Insumo) -> bool:
        return i.tipo_item == Insumo.TIPO_INTERNO or (i.codigo or "").startswith("DERIVADO:RECETA:")

    def _is_empaque(i: Insumo) -> bool:
        return i.tipo_item == Insumo.TIPO_EMPAQUE or "empaque" in normalizar_nombre(i.categoria or "")

    insumos_internos = [i for i in insumos if _is_internal(i)]
    insumos_empaque = [i for i in insumos if not _is_internal(i) and _is_empaque(i)]
    insumos_materia_prima = [i for i in insumos if not _is_internal(i) and not _is_empaque(i)]
    insumos_internos_ready_count = sum(1 for i in insumos_internos if i.erp_profile["ready"])
    insumos_empaque_ready_count = sum(1 for i in insumos_empaque if i.erp_profile["ready"])
    insumos_materia_prima_ready_count = sum(1 for i in insumos_materia_prima if i.erp_profile["ready"])
    quick_mode = receta.tipo == Receta.TIPO_PRODUCTO_FINAL and not advanced_mode and (
        linea is None or linea.tipo_linea == LineaReceta.TIPO_NORMAL
    )
    if component_filter not in {Insumo.TIPO_INTERNO, Insumo.TIPO_MATERIA_PRIMA, Insumo.TIPO_EMPAQUE, "ALL"}:
        component_filter = Insumo.TIPO_INTERNO if receta.tipo == Receta.TIPO_PRODUCTO_FINAL else Insumo.TIPO_MATERIA_PRIMA
    component_filter_meta = {
        "ALL": {
            "title": "Todo el catálogo",
            "description": "Muestra insumos internos, materia prima y empaques. Úsalo solo cuando no estés seguro de la clase.",
            "count": len(insumos),
        },
        Insumo.TIPO_INTERNO: {
            "title": "Insumos internos",
            "description": "Panes, rellenos, coberturas, batidas y subinsumos producidos dentro de la empresa.",
            "count": len(insumos_internos),
        },
        Insumo.TIPO_MATERIA_PRIMA: {
            "title": "Materia prima",
            "description": "Artículos comprados directo a proveedor y usados tal cual en la receta o batida.",
            "count": len(insumos_materia_prima),
        },
        Insumo.TIPO_EMPAQUE: {
            "title": "Empaques",
            "description": "Domo, caja, etiqueta, vaso y cualquier material de presentación final.",
            "count": len(insumos_empaque),
        },
    }
    component_context_map = {
        "internos": {
            "title": "Bloque: insumos internos",
            "description": "Estás agregando un componente producido dentro de la empresa: pan, relleno, cobertura, batida o subinsumo.",
        },
        "materia_prima": {
            "title": "Bloque: materia prima puntual",
            "description": "Estás agregando un artículo que entra directo al armado final sin pasar por una batida interna.",
        },
        "empaques": {
            "title": "Bloque: empaque final",
            "description": "Estás agregando caja, domo, etiqueta, vaso u otro material de presentación final.",
        },
        "subsecciones": {
            "title": "Bloque: subsección operativa",
            "description": "Estás capturando un bloque lógico como relleno, decorado o cobertura para ordenar el armado.",
        },
        "general": {
            "title": "Bloque general",
            "description": "Captura el componente dentro del bloque correcto para mantener el BOM ordenado.",
        },
    }
    if component_context not in component_context_map:
        component_context = "general"
    show_subsection_controls = component_context == "subsecciones" or (
        linea is not None and linea.tipo_linea == LineaReceta.TIPO_SUBSECCION
    )
    component_filter_choices = ["ALL", Insumo.TIPO_INTERNO, Insumo.TIPO_MATERIA_PRIMA, Insumo.TIPO_EMPAQUE]
    if quick_mode:
        component_filter_choices = [Insumo.TIPO_INTERNO, Insumo.TIPO_MATERIA_PRIMA, Insumo.TIPO_EMPAQUE]
    return {
        "receta": receta,
        "linea": linea,
        "advanced_mode": advanced_mode,
        "quick_mode": quick_mode,
        "default_component_filter": component_filter,
        "insumos": insumos,
        "insumos_internos": insumos_internos,
        "insumos_empaque": insumos_empaque,
        "insumos_materia_prima": insumos_materia_prima,
        "insumos_internos_count": len(insumos_internos),
        "insumos_empaque_count": len(insumos_empaque),
        "insumos_materia_prima_count": len(insumos_materia_prima),
        "insumos_internos_ready_count": insumos_internos_ready_count,
        "insumos_empaque_ready_count": insumos_empaque_ready_count,
        "insumos_materia_prima_ready_count": insumos_materia_prima_ready_count,
        "component_filter_meta": component_filter_meta,
        "component_context": component_context,
        "component_context_meta": component_context_map[component_context],
        "show_subsection_controls": show_subsection_controls,
        "component_filter_choices": component_filter_choices,
        "product_final_component_guidance": [
            {
                "filter": "INSUMO_INTERNO",
                "title": "Insumo interno",
                "description": "Usa panes, rellenos, coberturas y subinsumos ya producidos dentro de la empresa.",
            },
            {
                "filter": "MATERIA_PRIMA",
                "title": "Materia prima puntual",
                "description": "Úsala solo cuando el artículo entra directo al armado final y no existe como insumo interno.",
            },
            {
                "filter": "EMPAQUE",
                "title": "Empaque",
                "description": "Úsalo para domos, cajas, etiquetas, vasos y demás material de presentación final.",
            },
        ],
        "unidades": UnidadMedida.objects.order_by("codigo"),
        "linea_tipo_choices": LineaReceta.TIPO_CHOICES,
    }


def _latest_cost_for_insumo(insumo: Insumo | None) -> Decimal | None:
    if not insumo:
        return None
    cost, _unit, _source = resolve_insumo_unit_cost(insumo)
    return cost


def _switch_line_to_internal_cost(linea: LineaReceta) -> None:
    # Si ya hay cantidad + insumo, dejamos de usar costo fijo de Excel y
    # pasamos a costo dinámico interno (cantidad * costo_unitario_snapshot).
    if not linea.insumo:
        return
    if linea.cantidad is None or linea.cantidad <= 0:
        return

    if linea.costo_unitario_snapshot is None or linea.costo_unitario_snapshot <= 0:
        latest, _source = resolve_line_snapshot_cost(linea)
        if latest is not None and latest > 0:
            linea.costo_unitario_snapshot = latest
        elif linea.costo_linea_excel is not None and linea.costo_linea_excel > 0:
            try:
                linea.costo_unitario_snapshot = linea.costo_linea_excel / linea.cantidad
            except Exception:
                pass

    if linea.costo_unitario_snapshot is not None and linea.costo_unitario_snapshot > 0:
        linea.costo_linea_excel = None


def _autofill_unidad_from_insumo(linea: LineaReceta) -> None:
    if not linea.insumo:
        return
    if linea.unidad is None and linea.insumo.unidad_base is not None:
        linea.unidad = linea.insumo.unidad_base
    if linea.insumo.unidad_base is not None:
        # La unidad del componente ligado es fija para evitar inconsistencias de costeo.
        linea.unidad = linea.insumo.unidad_base
    if linea.unidad and not (linea.unidad_texto or "").strip():
        linea.unidad_texto = linea.unidad.codigo
    if linea.unidad:
        linea.unidad_texto = linea.unidad.codigo


def _validate_linea_operativa(receta: Receta, linea: LineaReceta) -> str | None:
    if linea.tipo_linea != LineaReceta.TIPO_NORMAL:
        return None
    if receta.tipo == Receta.TIPO_PRODUCTO_FINAL and not linea.insumo_id:
        return "En producto final, cada renglón principal debe estar ligado a un insumo/subinsumo."
    if linea.insumo_id and linea.insumo and linea.insumo.unidad_base_id is None:
        return (
            f"El insumo ligado ({linea.insumo.nombre}) no tiene unidad base. "
            "Asigna unidad en Catálogo de insumos para poder costear."
        )
    if linea.insumo_id and (linea.cantidad is None or linea.cantidad <= 0):
        unidad = linea.unidad.codigo if linea.unidad else (linea.unidad_texto or "").strip()
        if unidad:
            return f"Captura cantidad mayor a cero para el insumo ligado ({unidad})."
        return "Captura cantidad mayor a cero para el insumo ligado."

    # Regla operativa: en recetas tipo bollo, el pan se consume desde batida base (kg),
    # no desde presentaciones en pieza.
    presentacion = _extract_presentacion_from_recipe_name(receta.nombre)
    ingredient_text = ((linea.insumo_texto or "").strip() or (linea.insumo.nombre if linea.insumo else "")).lower()
    if presentacion in {"Bollo", "Bollos", "Bollito"} and ingredient_text.startswith("pan"):
        unit = linea.insumo.unidad_base if (linea.insumo and linea.insumo.unidad_base_id) else None
        if not unit or unit.tipo != UnidadMedida.TIPO_MASA:
            return (
                "En recetas de bollo, el pan debe ligarse a la batida base en kg. "
                "Selecciona el insumo derivado de preparación (no presentación por pieza)."
            )
    return None


def _line_component_kind(insumo: Insumo | None) -> str | None:
    if not insumo:
        return None
    if insumo.tipo_item == Insumo.TIPO_EMPAQUE:
        return Insumo.TIPO_EMPAQUE
    if insumo.tipo_item == Insumo.TIPO_INTERNO or (insumo.codigo or "").startswith("DERIVADO:RECETA:"):
        return Insumo.TIPO_INTERNO
    return Insumo.TIPO_MATERIA_PRIMA


def _validate_component_filter_selection(
    receta: Receta,
    linea: LineaReceta,
    component_filter: str | None,
) -> str | None:
    if receta.tipo != Receta.TIPO_PRODUCTO_FINAL:
        return None
    if not linea.insumo_id or not component_filter or component_filter == "ALL":
        return None
    expected = component_filter.strip().upper()
    if expected not in {Insumo.TIPO_INTERNO, Insumo.TIPO_MATERIA_PRIMA, Insumo.TIPO_EMPAQUE}:
        return None

    actual = _line_component_kind(linea.insumo)
    if actual == expected:
        return None

    expected_labels = {
        Insumo.TIPO_INTERNO: "insumo interno",
        Insumo.TIPO_MATERIA_PRIMA: "materia prima puntual",
        Insumo.TIPO_EMPAQUE: "empaque",
    }
    actual_labels = {
        Insumo.TIPO_INTERNO: "insumo interno",
        Insumo.TIPO_MATERIA_PRIMA: "materia prima",
        Insumo.TIPO_EMPAQUE: "empaque",
    }
    return (
        f"Seleccionaste el flujo de {expected_labels[expected]}, pero el artículo ligado "
        f"pertenece a {actual_labels.get(actual, 'otra clase')}. "
        "Cambia la clase de componente o selecciona un artículo del grupo correcto."
    )


def _validate_quick_mode_selection(
    receta: Receta,
    linea: LineaReceta,
    quick_mode: bool,
) -> str | None:
    if not quick_mode:
        return None
    if receta.tipo != Receta.TIPO_PRODUCTO_FINAL:
        return None
    if linea.tipo_linea != LineaReceta.TIPO_NORMAL:
        return None
    if not linea.insumo_id:
        return "En producto final, el modo rápido requiere seleccionar un artículo estándar del catálogo antes de guardar."
    latest_cost = _latest_cost_for_insumo(linea.insumo)
    if latest_cost is None or latest_cost <= 0:
        return (
            f"El artículo seleccionado ({linea.insumo.nombre}) no tiene costo vigente. "
            "Primero registra costo en el maestro para poder usarlo en producto final."
        )
    return None


def _repoint_linea_to_canonical_if_needed(linea: LineaReceta) -> bool:
    if not linea.insumo_id:
        return False
    canonical = _canonicalize_insumo_match(linea.insumo)
    if not canonical or canonical.id == linea.insumo_id:
        return False
    linea.insumo = canonical
    linea.insumo_texto = _insumo_display_name(canonical)[:250]
    return True


def _apply_direct_base_replacement_to_line(
    linea: LineaReceta,
    replacement: dict[str, object],
) -> None:
    linea.insumo = replacement["insumo"]
    linea.insumo_texto = _insumo_display_name(replacement["insumo"])
    linea.unidad = replacement["insumo"].unidad_base
    linea.unidad_texto = (
        replacement["insumo"].unidad_base.codigo
        if replacement["insumo"].unidad_base_id
        else linea.unidad_texto
    )
    linea.cantidad = replacement["replacement_quantity"]
    linea.costo_unitario_snapshot = None
    linea.costo_linea_excel = None
    linea.match_status = LineaReceta.STATUS_AUTO
    linea.match_method = "DIRECT_BASE_PRESENTACION"
    linea.match_score = 100


def _extract_presentacion_from_recipe_name(recipe_name: str) -> str | None:
    normalized = normalizar_nombre(recipe_name or "")
    if not normalized:
        return None
    token_map = [
        ("media plancha", "Media Plancha"),
        ("1 2 plancha", "Media Plancha"),
        ("rosca", "Rosca"),
        ("individual", "Individual"),
        ("bollito", "Bollito"),
        ("bollos", "Bollos"),
        ("bollo", "Bollos"),
        ("grande", "Grande"),
        ("mediano", "Mediano"),
        ("chico", "Chico"),
        ("mini", "Mini"),
    ]
    for token, label in token_map:
        if f" {token}" in f" {normalized}" or normalized.endswith(token):
            return label
    return None


def _meaningful_pan_tokens(*parts: str) -> list[str]:
    generic_tokens = {
        "pan",
        "de",
        "del",
        "la",
        "el",
        "los",
        "las",
        "con",
        "para",
        "y",
        "pastel",
        "pasteles",
        "producto",
        "final",
        "base",
        "batida",
        "mezcla",
        "relleno",
        "cobertura",
        "decorado",
        "decoracion",
        "chico",
        "mediano",
        "grande",
        "mini",
        "individual",
        "bollo",
        "bollos",
        "bollito",
        "rosca",
        "media",
        "plancha",
    }
    tokens: list[str] = []
    for part in parts:
        for token in normalizar_nombre(part or "").split():
            if len(token) < 3 or token in generic_tokens:
                continue
            if token not in tokens:
                tokens.append(token)
    return tokens


def _select_best_pan_derived_candidate(
    recipe_name: str,
    ingredient_text: str,
) -> Insumo | None:
    presentacion = _extract_presentacion_from_recipe_name(recipe_name)
    if not presentacion:
        return None

    ingredient_norm = normalizar_nombre(ingredient_text or "")
    recipe_norm = normalizar_nombre(recipe_name or "")
    full_context = f"{ingredient_norm} {recipe_norm}".strip()
    if not full_context.startswith("pan "):
        return None
    if full_context.startswith("decorado pan"):
        return None
    flavor_tokens = _meaningful_pan_tokens(ingredient_text, recipe_name)
    if not flavor_tokens:
        return None

    is_bollo_recipe = presentacion in {"Bollo", "Bollos", "Bollito"}
    presentacion_norm = normalizar_nombre(presentacion or "")
    candidates = list(
        Insumo.objects.filter(codigo__startswith="DERIVADO:RECETA:", activo=True, nombre__icontains="Pan")
        .only("id", "nombre", "codigo", "unidad_base_id")
        .order_by("nombre", "id")
    )
    best_candidate: Insumo | None = None
    best_score = -1
    for candidate in candidates:
        candidate_norm = normalizar_nombre(candidate.nombre or "")
        token_hits = [token for token in flavor_tokens if token in candidate_norm]
        if not token_hits:
            continue
        derived_kind = _derived_code_kind(candidate.codigo or "")
        score = len(token_hits) * 100
        if presentacion_norm and derived_kind == "PRESENTACION" and presentacion_norm in candidate_norm:
            score += 90
        if is_bollo_recipe and derived_kind == "PREPARACION":
            score += 140
        elif is_bollo_recipe and derived_kind == "PRESENTACION":
            score -= 120
        elif not is_bollo_recipe and derived_kind == "PRESENTACION":
            score += 50
        elif not is_bollo_recipe and derived_kind == "PREPARACION":
            score -= 40
        if candidate_norm.startswith("pan "):
            score += 10
        if score > best_score:
            best_score = score
            best_candidate = candidate
    return best_candidate if best_score >= 100 else None


def _autolink_pan_derived_from_recipe(receta: Receta, linea: LineaReceta) -> None:
    ingredient_text = (linea.insumo_texto or "").strip()
    if not ingredient_text and linea.insumo:
        ingredient_text = _insumo_display_name(linea.insumo)
    candidate = _select_best_pan_derived_candidate(receta.nombre, ingredient_text)
    if not candidate:
        return
    if linea.insumo_id == candidate.id:
        return
    linea.insumo = candidate


def _sync_derived_insumos_safe(request: HttpRequest, receta: Receta) -> None:
    try:
        sync_receta_derivados(receta)
    except Exception:
        messages.warning(
            request,
            "La receta se guardó, pero falló la sincronización automática de costos/insumos derivados.",
        )


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
def receta_sync_derivados(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    _sync_derived_insumos_safe(request, receta)
    _sync_cost_version_safe(request, receta, "SYNC_DERIVADOS_MANUAL")
    messages.success(request, "Derivados actualizados.")
    next_url = (request.GET.get("next") or "").strip()
    if next_url:
        return redirect(next_url)
    return redirect("recetas:receta_detail", pk=pk)


def _sync_cost_version_safe(request: HttpRequest, receta: Receta, fuente: str) -> None:
    try:
        asegurar_version_costeo(receta, fuente=fuente)
    except Exception:
        messages.warning(
            request,
            "Se guardaron cambios, pero falló el versionado automático de costos.",
        )


def _next_line_position(receta: Receta) -> int:
    last = receta.lineas.order_by("-posicion").values_list("posicion", flat=True).first()
    return int(last or 0) + 1


def _clone_linea_to_receta(
    source_linea: LineaReceta,
    target_receta: Receta,
    position: int,
) -> LineaReceta:
    return LineaReceta.objects.create(
        receta=target_receta,
        posicion=position,
        tipo_linea=source_linea.tipo_linea,
        etapa=source_linea.etapa,
        insumo=source_linea.insumo,
        insumo_texto=source_linea.insumo_texto,
        cantidad=source_linea.cantidad,
        unidad_texto=source_linea.unidad_texto,
        unidad=source_linea.unidad,
        costo_linea_excel=source_linea.costo_linea_excel,
        costo_unitario_snapshot=source_linea.costo_unitario_snapshot,
        match_score=source_linea.match_score,
        match_method=source_linea.match_method,
        match_status=source_linea.match_status,
        aprobado_por=source_linea.aprobado_por,
        aprobado_en=source_linea.aprobado_en,
    )


def _resequence_lineas(receta: Receta) -> None:
    changed: list[LineaReceta] = []
    for idx, linea in enumerate(receta.lineas.order_by("posicion", "id"), start=1):
        if linea.posicion != idx:
            linea.posicion = idx
            changed.append(linea)
    if changed:
        LineaReceta.objects.bulk_update(changed, ["posicion"])


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def receta_copy_lineas(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    source_id = (request.POST.get("source_receta_id") or "").strip()
    copy_mode = (request.POST.get("copy_mode") or "append").strip().lower()

    if not source_id.isdigit():
        messages.error(request, "Selecciona la receta origen que quieres copiar.")
        return redirect("recetas:receta_detail", pk=pk)

    receta_origen = get_object_or_404(Receta, pk=int(source_id))
    if receta_origen.pk == receta.pk:
        messages.error(request, "No puedes copiar ingredientes desde la misma receta.")
        return redirect("recetas:receta_detail", pk=pk)

    if copy_mode not in {"append", "replace"}:
        copy_mode = "append"

    lineas_origen = list(receta_origen.lineas.order_by("posicion", "id"))
    if not lineas_origen:
        messages.error(request, "La receta origen no tiene ingredientes para copiar.")
        return redirect("recetas:receta_detail", pk=pk)

    current_count = receta.lineas.count()
    deleted_count = 0
    copied_count = 0

    with transaction.atomic():
        if copy_mode == "replace":
            deleted_count, _ = receta.lineas.all().delete()
            start_position = 1
        else:
            start_position = _next_line_position(receta)

        position = start_position
        for source_linea in lineas_origen:
            _clone_linea_to_receta(source_linea, receta, position)
            position += 1
            copied_count += 1

        _resequence_lineas(receta)

    _sync_derived_insumos_safe(request, receta)
    _sync_cost_version_safe(request, receta, "LINEAS_COPY")
    log_event(
        request.user,
        "COPY_LINEAS",
        "recetas.Receta",
        str(receta.pk),
        {
            "receta_destino_id": receta.pk,
            "receta_destino_nombre": receta.nombre,
            "receta_origen_id": receta_origen.pk,
            "receta_origen_nombre": receta_origen.nombre,
            "copy_mode": copy_mode,
            "copied_count": copied_count,
            "deleted_count": deleted_count,
            "lineas_previas_destino": current_count,
        },
    )

    if copy_mode == "replace":
        messages.success(
            request,
            f"Se reemplazaron {deleted_count} líneas y se copiaron {copied_count} ingredientes desde {receta_origen.nombre}.",
        )
    else:
        messages.success(
            request,
            f"Se copiaron {copied_count} ingredientes desde {receta_origen.nombre} al final de la receta.",
        )
    return redirect("recetas:receta_detail", pk=pk)


def _load_versiones_costeo(receta: Receta, limit: int) -> list[RecetaCostoVersion]:
    return list(receta.versiones_costo.order_by("-version_num")[:limit])


def _empty_costeo_actual() -> dict[str, Any]:
    return {
        "driver": None,
        "costo_mp": Decimal("0"),
        "costo_mo": Decimal("0"),
        "costo_indirecto": Decimal("0"),
        "costo_total": Decimal("0"),
        "costo_por_unidad_rendimiento": Decimal("0"),
    }


def _compare_versions(base: RecetaCostoVersion, target: RecetaCostoVersion) -> dict[str, Decimal]:
    delta_mp = Decimal(str(target.costo_mp or 0)) - Decimal(str(base.costo_mp or 0))
    delta_mo = Decimal(str(target.costo_mo or 0)) - Decimal(str(base.costo_mo or 0))
    delta_ind = Decimal(str(target.costo_indirecto or 0)) - Decimal(str(base.costo_indirecto or 0))
    delta_total = Decimal(str(target.costo_total or 0)) - Decimal(str(base.costo_total or 0))

    base_total = Decimal(str(base.costo_total or 0))
    delta_pct_total = None
    if base_total > 0:
        delta_pct_total = (delta_total * Decimal("100")) / base_total

    base_unidad = Decimal(str(base.costo_por_unidad_rendimiento or 0))
    target_unidad = Decimal(str(target.costo_por_unidad_rendimiento or 0))
    delta_unidad = target_unidad - base_unidad
    delta_pct_unidad = None
    if base_unidad > 0:
        delta_pct_unidad = (delta_unidad * Decimal("100")) / base_unidad

    return {
        "delta_mp": delta_mp,
        "delta_mo": delta_mo,
        "delta_indirecto": delta_ind,
        "delta_total": delta_total,
        "delta_pct_total": delta_pct_total,
        "delta_unidad": delta_unidad,
        "delta_pct_unidad": delta_pct_unidad,
    }


def _map_driver_header(header: str) -> str:
    key = normalizar_nombre(header).replace("_", " ")
    if key in {"scope", "alcance", "tipo"}:
        return "scope"
    if key in {"nombre", "driver", "driver nombre"}:
        return "nombre"
    if key in {"receta", "producto", "nombre receta"}:
        return "receta"
    if key in {"codigo point", "codigo", "sku"}:
        return "codigo_point"
    if key in {"familia", "sheet", "categoria"}:
        return "familia"
    if key in {"lote desde", "lote min", "desde"}:
        return "lote_desde"
    if key in {"lote hasta", "lote max", "hasta"}:
        return "lote_hasta"
    if key in {"mo pct", "mo%", "mano obra pct", "mano de obra pct"}:
        return "mo_pct"
    if key in {"ind pct", "indirecto pct", "indirectos pct", "indirecto%"}:
        return "indirecto_pct"
    if key in {"mo fijo", "mano obra fijo", "mano de obra fijo"}:
        return "mo_fijo"
    if key in {"ind fijo", "indirecto fijo", "indirectos fijo"}:
        return "indirecto_fijo"
    if key in {"prioridad", "priority"}:
        return "prioridad"
    if key in {"activo", "enabled"}:
        return "activo"
    return key


def _normalize_driver_scope(raw: str | None) -> str:
    key = normalizar_nombre(raw or "")
    if key in {"producto", "product"}:
        return CostoDriver.SCOPE_PRODUCTO
    if key in {"familia", "family"}:
        return CostoDriver.SCOPE_FAMILIA
    if key in {"lote", "batch"}:
        return CostoDriver.SCOPE_LOTE
    return CostoDriver.SCOPE_GLOBAL


def _to_int_safe(value: object, default: int = 0) -> int:
    try:
        return int(str(value).strip())
    except Exception:
        return default


def _to_bool_safe(value: object, default: bool = True) -> bool:
    if value is None:
        return default
    key = normalizar_nombre(str(value))
    if key in {"1", "true", "si", "yes", "on", "activo"}:
        return True
    if key in {"0", "false", "no", "off", "inactivo"}:
        return False
    return default


def _load_driver_rows(uploaded) -> list[dict]:
    filename = (uploaded.name or "").lower()
    rows: list[dict] = []
    if filename.endswith(".csv"):
        uploaded.seek(0)
        content = uploaded.read().decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(content.splitlines())
        for row in reader:
            parsed = {}
            for key, value in (row or {}).items():
                if not key:
                    continue
                parsed[_map_driver_header(str(key))] = value
            rows.append(parsed)
        return rows
    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        uploaded.seek(0)
        wb = load_workbook(uploaded, read_only=True, data_only=True)
        ws = wb.active
        values = list(ws.values)
        if not values:
            return []
        headers = [_map_driver_header(str(h or "")) for h in values[0]]
        for raw in values[1:]:
            parsed = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                parsed[header] = raw[idx] if idx < len(raw) else None
            rows.append(parsed)
        return rows
    raise ValueError("Formato no soportado. Usa CSV o XLSX.")


@login_required
@permission_required("recetas.change_lineareceta", raise_exception=True)
def linea_edit(request: HttpRequest, pk: int, linea_id: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    linea = get_object_or_404(LineaReceta, pk=linea_id, receta=receta)

    if request.method == "POST":
        advanced_mode = receta.tipo == Receta.TIPO_PRODUCTO_FINAL and request.POST.get("advanced_mode") == "1"
        component_filter = (request.POST.get("component_filter") or "").strip().upper()
        component_context = (request.POST.get("component_context") or "").strip().lower()
        quick_mode = receta.tipo == Receta.TIPO_PRODUCTO_FINAL and not advanced_mode
        original_insumo_id = linea.insumo_id
        insumo_id = request.POST.get("insumo_id")
        selected_insumo = canonical_insumo_by_id(insumo_id)
        selected_was_variant = bool(selected_insumo and str(selected_insumo.id) != str(insumo_id or "").strip())
        tipo_linea = (request.POST.get("tipo_linea") or LineaReceta.TIPO_NORMAL).strip()
        if quick_mode:
            tipo_linea = LineaReceta.TIPO_NORMAL
        if tipo_linea not in {LineaReceta.TIPO_NORMAL, LineaReceta.TIPO_SUBSECCION}:
            tipo_linea = LineaReceta.TIPO_NORMAL
        linea.insumo_texto = (request.POST.get("insumo_texto") or "").strip()[:250]
        if not linea.insumo_texto and selected_insumo:
            linea.insumo_texto = _insumo_display_name(selected_insumo)[:250]
        linea.unidad_texto = (request.POST.get("unidad_texto") or "").strip()[:40]
        linea.etapa = (request.POST.get("etapa") or "").strip()[:120]
        linea.tipo_linea = tipo_linea
        linea.cantidad = _to_decimal_or_none(request.POST.get("cantidad"))
        linea.insumo = selected_insumo
        _autolink_pan_derived_from_recipe(receta, linea)
        canonicalized = _repoint_linea_to_canonical_if_needed(linea)
        if original_insumo_id != linea.insumo_id or canonicalized:
            # Si cambió el insumo ligado (manual/autolink), invalidamos snapshot previo
            # para recalcular costo unitario con el nuevo insumo.
            linea.costo_unitario_snapshot = None
            linea.costo_linea_excel = None
        if linea.insumo:
            linea.insumo_texto = _insumo_display_name(linea.insumo)[:250]

        if linea.insumo:
            linea.match_status = LineaReceta.STATUS_AUTO
            linea.match_method = "MANUAL"
            linea.match_score = 100.0
            linea.aprobado_por = request.user
            linea.aprobado_en = timezone.now()
        elif tipo_linea == LineaReceta.TIPO_SUBSECCION:
            linea.match_status = LineaReceta.STATUS_AUTO
            linea.match_method = LineaReceta.MATCH_SUBSECTION
            linea.match_score = 100.0
            linea.aprobado_por = request.user
            linea.aprobado_en = timezone.now()
        else:
            linea.match_status = LineaReceta.STATUS_REJECTED
            linea.match_method = LineaReceta.MATCH_NONE
            linea.match_score = 0.0

        _autofill_unidad_from_insumo(linea)
        validation_error = _validate_linea_operativa(receta, linea)
        validation_error = validation_error or _validate_component_filter_selection(receta, linea, component_filter)
        validation_error = validation_error or _validate_quick_mode_selection(receta, linea, quick_mode)
        validation_error = validation_error or _validate_selected_insumo_enterprise_ready(linea)
        if validation_error:
            messages.error(request, validation_error)
            return render(
                request,
                "recetas/linea_form.html",
                _linea_form_context(
                    receta,
                    linea,
                    advanced_mode=advanced_mode,
                    component_filter=component_filter,
                    component_context=component_context,
                ),
            )
        _switch_line_to_internal_cost(linea)
        linea.save()
        _sync_derived_insumos_safe(request, receta)
        _sync_cost_version_safe(request, receta, "LINEA_EDIT")
        if canonicalized or selected_was_variant:
            messages.info(request, f"El artículo se normalizó automáticamente a {_insumo_display_name(linea.insumo)}.")
        messages.success(request, "Línea actualizada.")
        return redirect("recetas:receta_detail", pk=pk)

    advanced_mode = receta.tipo == Receta.TIPO_PRODUCTO_FINAL and request.GET.get("advanced") == "1"
    component_filter = (request.GET.get("component_kind") or request.GET.get("component_filter") or "").strip().upper()
    component_context = (request.GET.get("component_context") or "").strip().lower()
    return render(
        request,
        "recetas/linea_form.html",
        _linea_form_context(
            receta,
            linea,
            advanced_mode=advanced_mode,
            component_filter=component_filter,
            component_context=component_context,
        ),
    )


@login_required
@permission_required("recetas.add_lineareceta", raise_exception=True)
def linea_create(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    if request.method == "POST":
        advanced_mode = receta.tipo == Receta.TIPO_PRODUCTO_FINAL and request.POST.get("advanced_mode") == "1"
        component_filter = (request.POST.get("component_filter") or "").strip().upper()
        component_context = (request.POST.get("component_context") or "").strip().lower()
        quick_mode = receta.tipo == Receta.TIPO_PRODUCTO_FINAL and not advanced_mode
        insumo_id = request.POST.get("insumo_id")
        selected_insumo = canonical_insumo_by_id(insumo_id)
        selected_was_variant = bool(selected_insumo and str(selected_insumo.id) != str(insumo_id or "").strip())
        tipo_linea = (request.POST.get("tipo_linea") or LineaReceta.TIPO_NORMAL).strip()
        if quick_mode:
            tipo_linea = LineaReceta.TIPO_NORMAL
        if tipo_linea not in {LineaReceta.TIPO_NORMAL, LineaReceta.TIPO_SUBSECCION}:
            tipo_linea = LineaReceta.TIPO_NORMAL
        posicion_default = (receta.lineas.order_by("-posicion").first().posicion + 1) if receta.lineas.exists() else 1
        insumo_texto = (request.POST.get("insumo_texto") or "").strip()[:250]
        if not insumo_texto and selected_insumo:
            insumo_texto = _insumo_display_name(selected_insumo)[:250]
        linea = LineaReceta(
            receta=receta,
            posicion=posicion_default,
            tipo_linea=tipo_linea,
            etapa=(request.POST.get("etapa") or "").strip()[:120],
            insumo_texto=insumo_texto,
            unidad_texto=(request.POST.get("unidad_texto") or "").strip()[:40],
            cantidad=_to_decimal_or_none(request.POST.get("cantidad")),
            costo_linea_excel=None,
            insumo=selected_insumo,
            unidad=None,
        )
        original_insumo_id = linea.insumo_id
        _autolink_pan_derived_from_recipe(receta, linea)
        canonicalized = _repoint_linea_to_canonical_if_needed(linea)
        if original_insumo_id != linea.insumo_id or canonicalized:
            linea.costo_unitario_snapshot = None
            linea.costo_linea_excel = None
        if linea.insumo:
            linea.insumo_texto = _insumo_display_name(linea.insumo)[:250]
        if linea.insumo:
            linea.match_status = LineaReceta.STATUS_AUTO
            linea.match_method = "MANUAL"
            linea.match_score = 100.0
            linea.aprobado_por = request.user
            linea.aprobado_en = timezone.now()
        elif tipo_linea == LineaReceta.TIPO_SUBSECCION:
            linea.match_status = LineaReceta.STATUS_AUTO
            linea.match_method = LineaReceta.MATCH_SUBSECTION
            linea.match_score = 100.0
            linea.aprobado_por = request.user
            linea.aprobado_en = timezone.now()
        else:
            linea.match_status = LineaReceta.STATUS_REJECTED
            linea.match_method = LineaReceta.MATCH_NONE
            linea.match_score = 0.0

        _autofill_unidad_from_insumo(linea)
        validation_error = _validate_linea_operativa(receta, linea)
        validation_error = validation_error or _validate_component_filter_selection(receta, linea, component_filter)
        validation_error = validation_error or _validate_quick_mode_selection(receta, linea, quick_mode)
        validation_error = validation_error or _validate_selected_insumo_enterprise_ready(linea)
        if validation_error:
            messages.error(request, validation_error)
            return render(
                request,
                "recetas/linea_form.html",
                _linea_form_context(
                    receta,
                    linea,
                    advanced_mode=advanced_mode,
                    component_filter=component_filter,
                    component_context=component_context,
                ),
            )
        _switch_line_to_internal_cost(linea)
        linea.save()
        _sync_derived_insumos_safe(request, receta)
        _sync_cost_version_safe(request, receta, "LINEA_CREATE")
        if canonicalized or selected_was_variant:
            messages.info(request, f"El artículo se normalizó automáticamente a {_insumo_display_name(linea.insumo)}.")
        messages.success(request, "Línea agregada.")
        return redirect("recetas:receta_detail", pk=pk)
    advanced_mode = receta.tipo == Receta.TIPO_PRODUCTO_FINAL and request.GET.get("advanced") == "1"
    component_filter = (request.GET.get("component_kind") or request.GET.get("component_filter") or "").strip().upper()
    component_context = (request.GET.get("component_context") or "").strip().lower()
    return render(
        request,
        "recetas/linea_form.html",
        _linea_form_context(
            receta,
            advanced_mode=advanced_mode,
            component_filter=component_filter,
            component_context=component_context,
        ),
    )


@login_required
@permission_required("recetas.delete_lineareceta", raise_exception=True)
@require_POST
def linea_delete(request: HttpRequest, pk: int, linea_id: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    linea = get_object_or_404(LineaReceta, pk=linea_id, receta=receta)
    linea.delete()
    _sync_derived_insumos_safe(request, receta)
    _sync_cost_version_safe(request, receta, "LINEA_DELETE")
    messages.success(request, "Línea eliminada.")
    return redirect("recetas:receta_detail", pk=pk)


@login_required
@permission_required("recetas.change_lineareceta", raise_exception=True)
@require_POST
def linea_apply_direct_base_replacement(request: HttpRequest, pk: int, linea_id: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    linea = get_object_or_404(LineaReceta.objects.select_related("insumo", "unidad"), pk=linea_id, receta=receta)
    linea.source_recipe = None
    linea.source_code_kind = None
    linea.source_active_presentaciones_count = 0
    linea.uses_direct_base_in_final = False
    if linea.insumo_id:
        linea.source_code_kind = _derived_code_kind(linea.insumo.codigo or "")
        source_recipe_id = _recipe_id_from_derived_code(linea.insumo.codigo or "")
        if source_recipe_id:
            linea.source_recipe = (
                Receta.objects.filter(pk=source_recipe_id).only("id", "nombre", "usa_presentaciones").first()
            )
            linea.source_active_presentaciones_count = RecetaPresentacion.objects.filter(
                receta_id=source_recipe_id,
                activo=True,
            ).count()
            direct_base_candidate = bool(
                receta.tipo == Receta.TIPO_PRODUCTO_FINAL
                and linea.insumo.tipo_item == Insumo.TIPO_INTERNO
                and linea.source_code_kind == "PREPARACION"
                and linea.source_recipe
                and linea.source_recipe.usa_presentaciones
                and linea.source_active_presentaciones_count > 0
            )
            if direct_base_candidate:
                direct_base_candidates = _active_presentation_derived_candidates(source_recipe_id)
                linea.uses_direct_base_in_final = bool(_suggest_direct_base_replacement(linea)) or not direct_base_candidates

    replacement = _suggest_direct_base_replacement(linea)
    if not replacement:
        messages.warning(
            request,
            "No se encontró una presentación derivada sugerida para esta línea.",
        )
        return redirect("recetas:receta_detail", pk=pk)

    original_insumo = linea.insumo
    original_cantidad = Decimal(str(linea.cantidad or 0))
    _apply_direct_base_replacement_to_line(linea, replacement)
    linea.save(
        update_fields=[
            "insumo",
            "insumo_texto",
            "unidad",
            "unidad_texto",
            "cantidad",
            "costo_unitario_snapshot",
            "costo_linea_excel",
            "match_status",
            "match_method",
            "match_score",
        ]
    )
    _sync_cost_version_safe(request, receta, "LINEA_DIRECT_BASE_REPLACEMENT")
    log_event(
        request.user,
        "LINEA_DIRECT_BASE_REPLACEMENT",
        "recetas.LineaReceta",
        str(linea.id),
        {
            "receta_id": receta.id,
            "linea_id": linea.id,
            "from_insumo_id": original_insumo.id if original_insumo else None,
            "to_insumo_id": replacement["insumo"].id,
            "from_cantidad": str(original_cantidad),
            "to_cantidad": str(replacement["replacement_quantity"]),
            "presentacion_id": replacement["presentacion"].id,
        },
    )
    messages.success(
        request,
        f"Línea actualizada a {replacement['insumo'].nombre} ({replacement['replacement_quantity']} {linea.unidad_texto}).",
    )
    return redirect("recetas:receta_detail", pk=pk)


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
def presentacion_create(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    workflow_errors = _validate_presentacion_workflow(receta)
    if workflow_errors:
        for error in workflow_errors:
            messages.error(request, error)
        return redirect("recetas:receta_detail", pk=pk)
    if request.method == "POST":
        nombre = (request.POST.get("nombre") or "").strip()
        peso_por_unidad_kg = _to_decimal_or_none(request.POST.get("peso_por_unidad_kg"))
        activo = request.POST.get("activo") == "on"
        if not nombre:
            messages.error(request, "El nombre de la presentación es obligatorio.")
            return redirect("recetas:presentacion_create", pk=pk)
        if not peso_por_unidad_kg or peso_por_unidad_kg <= 0:
            messages.error(request, "Cantidad por presentación debe ser mayor que cero.")
            return redirect("recetas:presentacion_create", pk=pk)

        presentacion, _ = RecetaPresentacion.objects.update_or_create(
            receta=receta,
            nombre=nombre[:80],
            defaults={
                "peso_por_unidad_kg": peso_por_unidad_kg,
                "activo": activo,
            },
        )
        if not receta.usa_presentaciones:
            receta.usa_presentaciones = True
            receta.save(update_fields=["usa_presentaciones"])
        _sync_derived_insumos_safe(request, receta)
        _sync_cost_version_safe(request, receta, "PRESENTACION_CREATE")
        messages.success(request, "Presentación guardada.")
        return redirect("recetas:receta_detail", pk=pk)

    return render(
        request,
        "recetas/presentacion_form.html",
        {
            "receta": receta,
            "presentacion": None,
            "rendimiento_unit_code": (receta.rendimiento_unidad.codigo if receta.rendimiento_unidad else ""),
        },
    )


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
def presentacion_edit(request: HttpRequest, pk: int, presentacion_id: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    workflow_errors = _validate_presentacion_workflow(receta)
    if workflow_errors:
        for error in workflow_errors:
            messages.error(request, error)
        return redirect("recetas:receta_detail", pk=pk)
    presentacion = get_object_or_404(RecetaPresentacion, pk=presentacion_id, receta=receta)
    if request.method == "POST":
        nombre = (request.POST.get("nombre") or "").strip()
        peso_por_unidad_kg = _to_decimal_or_none(request.POST.get("peso_por_unidad_kg"))
        activo = request.POST.get("activo") == "on"
        if not nombre:
            messages.error(request, "El nombre de la presentación es obligatorio.")
            return redirect("recetas:presentacion_edit", pk=pk, presentacion_id=presentacion_id)
        if not peso_por_unidad_kg or peso_por_unidad_kg <= 0:
            messages.error(request, "Cantidad por presentación debe ser mayor que cero.")
            return redirect("recetas:presentacion_edit", pk=pk, presentacion_id=presentacion_id)

        presentacion.nombre = nombre[:80]
        presentacion.peso_por_unidad_kg = peso_por_unidad_kg
        presentacion.activo = activo
        presentacion.save()
        _sync_derived_insumos_safe(request, receta)
        _sync_cost_version_safe(request, receta, "PRESENTACION_EDIT")
        messages.success(request, "Presentación actualizada.")
        return redirect("recetas:receta_detail", pk=pk)

    return render(
        request,
        "recetas/presentacion_form.html",
        {
            "receta": receta,
            "presentacion": presentacion,
            "rendimiento_unit_code": (receta.rendimiento_unidad.codigo if receta.rendimiento_unidad else ""),
        },
    )


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def presentacion_delete(request: HttpRequest, pk: int, presentacion_id: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    workflow_errors = _validate_presentacion_workflow(receta)
    if workflow_errors:
        for error in workflow_errors:
            messages.error(request, error)
        return redirect("recetas:receta_detail", pk=pk)
    presentacion = get_object_or_404(RecetaPresentacion, pk=presentacion_id, receta=receta)
    try:
        sync_presentacion_insumo(presentacion, deactivate=True)
    except Exception:
        messages.warning(
            request,
            "La presentación se eliminó, pero falló la desactivación del insumo derivado.",
        )
    presentacion.delete()
    _sync_cost_version_safe(request, receta, "PRESENTACION_DELETE")
    messages.success(request, "Presentación eliminada.")
    return redirect("recetas:receta_detail", pk=pk)


@login_required
@permission_required("recetas.view_receta", raise_exception=True)
def receta_versiones_export(request: HttpRequest, pk: int) -> HttpResponse:
    receta = get_object_or_404(Receta, pk=pk)
    export_format = (request.GET.get("format") or "csv").strip().lower()
    try:
        versiones = _load_versiones_costeo(receta, 300)
    except (OperationalError, ProgrammingError):
        versiones = []

    if export_format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "versiones_costeo"
        ws.append(
            [
                "version",
                "fecha",
                "fuente",
                "driver_scope",
                "driver_nombre",
                "costo_mp",
                "costo_mo",
                "costo_indirecto",
                "costo_total",
                "rendimiento",
                "unidad_rendimiento",
                "costo_por_unidad_rendimiento",
            ]
        )
        for v in versiones:
            ws.append(
                [
                    v.version_num,
                    timezone.localtime(v.creado_en).strftime("%Y-%m-%d %H:%M"),
                    v.fuente,
                    v.driver_scope,
                    v.driver_nombre,
                    float(v.costo_mp or 0),
                    float(v.costo_mo or 0),
                    float(v.costo_indirecto or 0),
                    float(v.costo_total or 0),
                    float(v.rendimiento_cantidad or 0),
                    v.rendimiento_unidad,
                    float(v.costo_por_unidad_rendimiento or 0),
                ]
            )
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        resp = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="receta_{receta.id}_versiones_costeo.xlsx"'
        return resp

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="receta_{receta.id}_versiones_costeo.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "version",
            "fecha",
            "fuente",
            "driver_scope",
            "driver_nombre",
            "costo_mp",
            "costo_mo",
            "costo_indirecto",
            "costo_total",
            "rendimiento",
            "unidad_rendimiento",
            "costo_por_unidad_rendimiento",
        ]
    )
    for v in versiones:
        writer.writerow(
            [
                v.version_num,
                timezone.localtime(v.creado_en).strftime("%Y-%m-%d %H:%M"),
                v.fuente,
                v.driver_scope,
                v.driver_nombre,
                v.costo_mp,
                v.costo_mo,
                v.costo_indirecto,
                v.costo_total,
                v.rendimiento_cantidad,
                v.rendimiento_unidad,
                v.costo_por_unidad_rendimiento,
            ]
        )
    return response


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
def drivers_costeo(request: HttpRequest) -> HttpResponse:
    q = (request.GET.get("q") or "").strip()
    scope = (request.GET.get("scope") or "").strip().upper()
    edit_raw = (request.GET.get("edit") or "").strip()
    edit_driver = None

    drivers_unavailable = False
    drivers = []
    try:
        qs = CostoDriver.objects.select_related("receta").order_by("scope", "prioridad", "id")
        if q:
            qs = qs.filter(
                Q(nombre__icontains=q)
                | Q(familia__icontains=q)
                | Q(receta__nombre__icontains=q)
            )
        if scope in {CostoDriver.SCOPE_PRODUCTO, CostoDriver.SCOPE_FAMILIA, CostoDriver.SCOPE_LOTE, CostoDriver.SCOPE_GLOBAL}:
            qs = qs.filter(scope=scope)
        if edit_raw:
            try:
                edit_driver = CostoDriver.objects.select_related("receta").filter(pk=int(edit_raw)).first()
            except Exception:
                edit_driver = None
        drivers = list(qs[:400])
    except (OperationalError, ProgrammingError):
        drivers_unavailable = True
        edit_driver = None

    if request.method == "POST":
        if drivers_unavailable:
            messages.error(request, "Drivers de costeo no disponibles en este entorno. Ejecuta migraciones.")
            return redirect("recetas:drivers_costeo")
        driver_id = (request.POST.get("driver_id") or "").strip()
        nombre = (request.POST.get("nombre") or "").strip()
        driver_scope = _normalize_driver_scope(request.POST.get("scope"))
        receta_id = (request.POST.get("receta_id") or "").strip()
        familia = (request.POST.get("familia") or "").strip()
        lote_desde = _to_decimal_or_none(request.POST.get("lote_desde"))
        lote_hasta = _to_decimal_or_none(request.POST.get("lote_hasta"))
        mo_pct = _to_decimal_safe(request.POST.get("mo_pct"))
        indirecto_pct = _to_decimal_safe(request.POST.get("indirecto_pct"))
        mo_fijo = _to_decimal_safe(request.POST.get("mo_fijo"))
        indirecto_fijo = _to_decimal_safe(request.POST.get("indirecto_fijo"))
        prioridad = _to_int_safe(request.POST.get("prioridad"), 100)
        activo = _to_bool_safe(request.POST.get("activo"), True)

        if not nombre:
            messages.error(request, "Nombre del driver es obligatorio.")
            return redirect("recetas:drivers_costeo")

        receta = Receta.objects.filter(pk=receta_id).first() if receta_id else None
        if driver_scope == CostoDriver.SCOPE_PRODUCTO and not receta:
            messages.error(request, "Para scope PRODUCTO debes seleccionar una receta.")
            return redirect("recetas:drivers_costeo")

        try:
            if driver_id:
                driver = CostoDriver.objects.filter(pk=driver_id).first()
                if not driver:
                    messages.error(request, "Driver no encontrado para editar.")
                    return redirect("recetas:drivers_costeo")
            else:
                driver = CostoDriver()

            driver.nombre = nombre[:120]
            driver.scope = driver_scope
            driver.receta = receta
            driver.familia = familia[:120]
            driver.lote_desde = lote_desde
            driver.lote_hasta = lote_hasta
            driver.mo_pct = mo_pct
            driver.indirecto_pct = indirecto_pct
            driver.mo_fijo = mo_fijo
            driver.indirecto_fijo = indirecto_fijo
            driver.prioridad = max(prioridad, 0)
            driver.activo = activo
            driver.save()
            messages.success(request, "Driver guardado.")
        except (OperationalError, ProgrammingError):
            messages.error(request, "No se pudo guardar el driver. Ejecuta migraciones pendientes.")
        return redirect("recetas:drivers_costeo")

    return render(
        request,
        "recetas/drivers_costeo.html",
        {
            "drivers": drivers,
            "recetas": Receta.objects.order_by("nombre"),
            "q": q,
            "scope": scope,
            "edit_driver": edit_driver,
            "drivers_unavailable": drivers_unavailable,
            "scope_choices": [
                CostoDriver.SCOPE_PRODUCTO,
                CostoDriver.SCOPE_FAMILIA,
                CostoDriver.SCOPE_LOTE,
                CostoDriver.SCOPE_GLOBAL,
            ],
        },
    )


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def drivers_costeo_delete(request: HttpRequest, driver_id: int) -> HttpResponse:
    try:
        driver = get_object_or_404(CostoDriver, pk=driver_id)
        driver.delete()
        messages.success(request, "Driver eliminado.")
    except (OperationalError, ProgrammingError):
        messages.error(request, "No se pudo eliminar el driver. Ejecuta migraciones pendientes.")
    return redirect("recetas:drivers_costeo")


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
def drivers_costeo_plantilla(request: HttpRequest) -> HttpResponse:
    export_format = (request.GET.get("format") or "xlsx").strip().lower()
    headers = [
        "scope",
        "nombre",
        "receta",
        "codigo_point",
        "familia",
        "lote_desde",
        "lote_hasta",
        "mo_pct",
        "indirecto_pct",
        "mo_fijo",
        "indirecto_fijo",
        "prioridad",
        "activo",
    ]
    rows = [
        ["PRODUCTO", "Driver Pastel Fresas", "Pastel Fresas Con Crema - Chico", "PFC-CHICO", "", "", "", "8", "4", "0", "0", "10", "1"],
        ["FAMILIA", "Driver Insumos 1", "", "", "Insumos 1", "", "", "6", "3", "0", "0", "30", "1"],
        ["GLOBAL", "Driver Global Base", "", "", "", "", "", "5", "2", "0", "0", "100", "1"],
    ]

    if export_format == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="plantilla_drivers_costeo.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response

    wb = Workbook()
    ws = wb.active
    ws.title = "drivers_costeo"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="F8E8EE")
    header_font = Font(color="8B2252", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.auto_filter.ref = f"A1:M{len(rows) + 1}"
    ws.freeze_panes = "A2"

    widths = {
        "A": 14,   # scope
        "B": 30,   # nombre
        "C": 32,   # receta
        "D": 20,   # codigo_point
        "E": 20,   # familia
        "F": 12,   # lote_desde
        "G": 12,   # lote_hasta
        "H": 10,   # mo_pct
        "I": 12,   # indirecto_pct
        "J": 10,   # mo_fijo
        "K": 14,   # indirecto_fijo
        "L": 10,   # prioridad
        "M": 8,    # activo
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws_guide = wb.create_sheet("instrucciones")
    ws_guide.append(["Campo", "Obligatorio", "Descripción / valores válidos"])
    guide_rows = [
        ["scope", "Sí", "GLOBAL | FAMILIA | PRODUCTO | LOTE"],
        ["nombre", "Sí", "Nombre descriptivo del driver."],
        ["receta", "Condicional", "Obligatorio si scope=PRODUCTO."],
        ["codigo_point", "Opcional", "Alias de receta para encontrarla por código comercial."],
        ["familia", "Condicional", "Obligatorio si scope=FAMILIA."],
        ["lote_desde", "Condicional", "Usar con scope=LOTE (decimal >= 0)."],
        ["lote_hasta", "Condicional", "Usar con scope=LOTE (decimal >= lote_desde)."],
        ["mo_pct", "Sí", "Porcentaje de mano de obra, por ejemplo 8."],
        ["indirecto_pct", "Sí", "Porcentaje de indirectos, por ejemplo 4."],
        ["mo_fijo", "Opcional", "Monto fijo adicional de MO."],
        ["indirecto_fijo", "Opcional", "Monto fijo adicional de indirectos."],
        ["prioridad", "Sí", "Menor número = mayor prioridad."],
        ["activo", "Sí", "1 activo / 0 inactivo."],
    ]
    for row in guide_rows:
        ws_guide.append(row)
    for col in range(1, 4):
        cell = ws_guide.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    ws_guide.column_dimensions["A"].width = 22
    ws_guide.column_dimensions["B"].width = 14
    ws_guide.column_dimensions["C"].width = 72
    ws_guide.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_drivers_costeo.xlsx"'
    return response


@login_required
@permission_required("recetas.change_receta", raise_exception=True)
@require_POST
def drivers_costeo_importar(request: HttpRequest) -> HttpResponse:
    try:
        CostoDriver.objects.exists()
    except (OperationalError, ProgrammingError):
        messages.error(request, "Drivers de costeo no disponibles en este entorno. Ejecuta migraciones.")
        return redirect("recetas:drivers_costeo")

    uploaded = request.FILES.get("archivo")
    modo = (request.POST.get("modo") or "upsert").strip().lower()
    if modo not in {"upsert", "replace"}:
        modo = "upsert"

    if not uploaded:
        messages.error(request, "Selecciona un archivo para importar drivers.")
        return redirect("recetas:drivers_costeo")

    try:
        rows = _load_driver_rows(uploaded)
    except Exception as exc:
        messages.error(request, f"No se pudo leer archivo de drivers: {exc}")
        return redirect("recetas:drivers_costeo")

    if not rows:
        messages.warning(request, "Archivo de drivers sin filas.")
        return redirect("recetas:drivers_costeo")

    created = 0
    updated = 0
    skipped = 0

    if modo == "replace":
        CostoDriver.objects.all().delete()

    for row in rows:
        scope = _normalize_driver_scope(row.get("scope"))
        nombre = str(row.get("nombre") or "").strip()
        if not nombre:
            skipped += 1
            continue

        receta = None
        receta_name = str(row.get("receta") or "").strip()
        codigo_point = str(row.get("codigo_point") or "").strip()
        if codigo_point:
            receta = Receta.objects.filter(codigo_point__iexact=codigo_point).order_by("id").first()
        if receta is None and receta_name:
            receta = Receta.objects.filter(nombre_normalizado=normalizar_nombre(receta_name)).order_by("id").first()

        if scope == CostoDriver.SCOPE_PRODUCTO and not receta:
            skipped += 1
            continue

        familia = str(row.get("familia") or "").strip()
        lote_desde = _to_decimal_or_none(str(row.get("lote_desde") or "").strip())
        lote_hasta = _to_decimal_or_none(str(row.get("lote_hasta") or "").strip())

        filter_kwargs = {
            "scope": scope,
            "nombre": nombre[:120],
            "receta": receta,
            "familia_normalizada": normalizar_nombre(familia),
            "lote_desde": lote_desde,
            "lote_hasta": lote_hasta,
        }
        driver = CostoDriver.objects.filter(**filter_kwargs).first()
        if not driver:
            driver = CostoDriver(**filter_kwargs)
            created += 1
        else:
            updated += 1

        driver.familia = familia[:120]
        driver.mo_pct = _to_decimal_safe(row.get("mo_pct"))
        driver.indirecto_pct = _to_decimal_safe(row.get("indirecto_pct"))
        driver.mo_fijo = _to_decimal_safe(row.get("mo_fijo"))
        driver.indirecto_fijo = _to_decimal_safe(row.get("indirecto_fijo"))
        driver.prioridad = max(_to_int_safe(row.get("prioridad"), 100), 0)
        driver.activo = _to_bool_safe(row.get("activo"), True)
        driver.save()

    messages.success(
        request,
        f"Importación drivers completada. Creados: {created}. Actualizados: {updated}. Omitidos: {skipped}.",
    )
    return redirect("recetas:drivers_costeo")
