from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from recetas.utils.normalizacion import normalizar_nombre


def _clean(value: Any) -> str:
    if value is None:
        return ""
    txt = str(value).strip()
    if txt.lower() in {"none", "nan"}:
        return ""
    return txt


class Command(BaseCommand):
    help = "Genera plantillas de captura de reabasto (una por sucursal) desde consolidado de stock mínimo."

    def add_arguments(self, parser):
        parser.add_argument(
            "archivo",
            type=str,
            help="Ruta al XLSX consolidado generado por extraer_stock_minimos_sucursales.",
        )
        parser.add_argument(
            "--sheet",
            type=str,
            default="stock_minimo_abastecimiento",
            help="Hoja a utilizar del consolidado.",
        )
        parser.add_argument(
            "--output-dir",
            type=str,
            default="output/spreadsheet/plantillas_sucursales",
            help="Carpeta donde se crearán las plantillas por sucursal.",
        )
        parser.add_argument(
            "--periodo",
            type=str,
            default="",
            help="Etiqueta de periodo para nombre de archivo (ej. 2026-02).",
        )

    def handle(self, *args, **options):
        input_path = Path(options["archivo"]).expanduser()
        if not input_path.exists():
            raise CommandError(f"No existe archivo: {input_path}")
        if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise CommandError("Formato no soportado. Usa .xlsx/.xlsm.")

        sheet_name = (options.get("sheet") or "stock_minimo_abastecimiento").strip()
        out_dir = Path(options.get("output_dir") or "output/spreadsheet/plantillas_sucursales")
        out_dir.mkdir(parents=True, exist_ok=True)

        periodo_raw = (options.get("periodo") or "").strip()
        if not periodo_raw:
            now = datetime.now()
            periodo_raw = f"{now.year:04d}-{now.month:02d}"
        periodo_safe = periodo_raw.replace("/", "-").replace(" ", "_")

        wb = load_workbook(input_path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"No existe hoja '{sheet_name}' en {input_path.name}.")
        ws = wb[sheet_name]

        data = list(ws.iter_rows(values_only=True))
        if not data:
            raise CommandError("Hoja sin datos.")

        headers = [normalizar_nombre(str(h or "")).replace("_", " ") for h in data[0]]
        idx = {h: i for i, h in enumerate(headers) if h}
        required = {
            "sucursal",
            "periodo",
            "categoria",
            "producto",
            "presentacion",
            "stock minimo",
            "receta match",
            "codigo point",
            "match status",
        }
        if not required.issubset(set(idx)):
            raise CommandError(
                "El consolidado no tiene columnas esperadas (sucursal, periodo, categoria, producto, presentacion, stock minimo, receta match, codigo point, match status)."
            )

        grouped: dict[str, list[dict[str, Any]]] = {}
        for raw in data[1:]:
            suc = _clean(raw[idx["sucursal"]])
            if not suc:
                continue

            categoria = _clean(raw[idx["categoria"]])
            producto = _clean(raw[idx["producto"]])
            presentacion = _clean(raw[idx["presentacion"]])
            if not producto:
                continue

            producto_display = producto
            if presentacion:
                producto_display = f"{producto} - {presentacion}"

            grouped.setdefault(suc, []).append(
                {
                    "periodo": _clean(raw[idx["periodo"]]).upper(),
                    "categoria": categoria,
                    "producto": producto,
                    "presentacion": presentacion,
                    "producto_display": producto_display,
                    "stock_minimo": raw[idx["stock minimo"]],
                    "receta_match": _clean(raw[idx["receta match"]]),
                    "codigo_point": _clean(raw[idx["codigo point"]]),
                    "match_status": _clean(raw[idx["match status"]]).upper(),
                }
            )

        if not grouped:
            raise CommandError("No hay filas de sucursal para generar plantillas.")

        headers_out = [
            "fecha_operacion",
            "sucursal",
            "periodo",
            "categoria",
            "producto",
            "presentacion",
            "codigo_point",
            "receta",
            "match_status",
            "stock_minimo",
            "stock_final_cierre",
            "solicitud_requerida",
            "solicitud_final",
            "observaciones",
        ]

        fill_header = PatternFill(start_color="F8E8EE", end_color="F8E8EE", fill_type="solid")
        fill_hint = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        font_header = Font(bold=True)

        created_files: list[Path] = []
        for sucursal, items in sorted(grouped.items(), key=lambda x: x[0].lower()):
            items_sorted = sorted(items, key=lambda x: (x["periodo"], normalizar_nombre(x["producto_display"])))
            # dedupe by producto+presentacion+periodo taking highest stock
            dedupe: dict[tuple[str, str], dict[str, Any]] = {}
            for it in items_sorted:
                key = (it["periodo"], normalizar_nombre(it["producto_display"]))
                prev = dedupe.get(key)
                current_stock = float(it["stock_minimo"] or 0)
                prev_stock = float(prev["stock_minimo"] or 0) if prev else -1
                if prev is None or current_stock > prev_stock:
                    dedupe[key] = it
            items_sorted = sorted(dedupe.values(), key=lambda x: (x["periodo"], normalizar_nombre(x["producto_display"])))

            out_wb = Workbook()
            out_ws = out_wb.active
            out_ws.title = "captura_reabasto"
            out_ws.append(headers_out)
            for col in range(1, len(headers_out) + 1):
                cell = out_ws.cell(1, col)
                cell.fill = fill_header
                cell.font = font_header

            for it in items_sorted:
                next_row = out_ws.max_row + 1
                stock_cell = f"J{next_row}"
                cierre_cell = f"K{next_row}"
                out_ws.append(
                    [
                        "",  # fecha_operacion
                        sucursal,
                        it["periodo"],
                        it["categoria"],
                        it["producto"],
                        it["presentacion"],
                        it["codigo_point"],
                        it["receta_match"],
                        it["match_status"],
                        it["stock_minimo"],
                        "",  # stock_final_cierre
                        f"=MAX({stock_cell}-IFERROR({cierre_cell},0),0)",  # solicitud_requerida
                        "",  # solicitud_final
                        "",  # observaciones
                    ]
                )

            # Highlight user editable columns
            for row in range(2, out_ws.max_row + 1):
                for col in (1, 11, 13, 14):
                    out_ws.cell(row, col).fill = fill_hint

            out_ws.freeze_panes = "A2"
            out_ws.auto_filter.ref = f"A1:N{max(2, out_ws.max_row)}"
            out_ws.column_dimensions["A"].width = 14
            out_ws.column_dimensions["B"].width = 18
            out_ws.column_dimensions["C"].width = 10
            out_ws.column_dimensions["D"].width = 12
            out_ws.column_dimensions["E"].width = 34
            out_ws.column_dimensions["F"].width = 15
            out_ws.column_dimensions["G"].width = 14
            out_ws.column_dimensions["H"].width = 32
            out_ws.column_dimensions["I"].width = 12
            out_ws.column_dimensions["J"].width = 12
            out_ws.column_dimensions["K"].width = 14
            out_ws.column_dimensions["L"].width = 16
            out_ws.column_dimensions["M"].width = 14
            out_ws.column_dimensions["N"].width = 24

            safe_name = normalizar_nombre(sucursal).replace("_", "-").replace(" ", "-")
            output_file = out_dir / f"plantilla_reabasto_{safe_name}_{periodo_safe}.xlsx"
            out_wb.save(output_file)
            created_files.append(output_file)

        # Also create quick guide
        guide = out_dir / f"README_plantillas_reabasto_{periodo_safe}.txt"
        guide.write_text(
            "\n".join(
                [
                    f"Plantillas reabasto periodo {periodo_raw}",
                    "",
                    "Columnas que SI llena sucursal:",
                    "- fecha_operacion (YYYY-MM-DD)",
                    "- stock_final_cierre",
                    "- solicitud_final (si la sucursal quiere ajustar manualmente)",
                    "- observaciones",
                    "",
                    "Columnas de referencia (no editar):",
                    "- sucursal, periodo, categoria, producto, presentacion, receta, codigo_point, match_status, stock_minimo",
                    "",
                    "Regla:",
                    "- solicitud_requerida = MAX(stock_minimo - stock_final_cierre, 0)",
                ]
            ),
            encoding="utf-8",
        )

        self.stdout.write(self.style.SUCCESS("Plantillas de sucursal generadas"))
        self.stdout.write(f"  - archivo fuente: {input_path}")
        self.stdout.write(f"  - hoja: {sheet_name}")
        self.stdout.write(f"  - sucursales: {len(created_files)}")
        self.stdout.write(f"  - carpeta salida: {out_dir}")
        for f in created_files:
            self.stdout.write(f"    * {f}")
        self.stdout.write(f"  - guía: {guide}")
