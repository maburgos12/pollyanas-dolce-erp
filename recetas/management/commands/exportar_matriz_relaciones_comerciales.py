from __future__ import annotations

import csv
from dataclasses import asdict
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from recetas.utils.commercial_composition import iter_commercial_validation_rows


class Command(BaseCommand):
    help = "Exporta la matriz maestra de relaciones comerciales Point-ERP a XLSX y CSV."

    def handle(self, *args, **options):
        rows = iter_commercial_validation_rows()
        generated_at = timezone.localtime()
        date_label = generated_at.date().isoformat()

        output_dir = Path(settings.BASE_DIR) / "output" / "spreadsheet" / "validacion_negocio"
        output_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = output_dir / f"matriz_validacion_relaciones_point_erp_{date_label}.xlsx"
        csv_path = output_dir / f"matriz_validacion_relaciones_point_erp_{date_label}.csv"

        self._write_workbook(xlsx_path, rows, generated_at.isoformat())
        self._write_csv(csv_path, rows)

        self.stdout.write(str(xlsx_path))
        self.stdout.write(str(csv_path))
        self.stdout.write(
            str(
                {
                    "total": len(rows),
                    "historico": sum(1 for row in rows if row.clasificacion == "HISTORICO_LEGADO"),
                    "complemento": sum(1 for row in rows if row.clasificacion == "COMPLEMENTO_OBLIGATORIO"),
                    "directo": sum(1 for row in rows if row.clasificacion == "PRODUCTO_BASE_DIRECTO"),
                    "sin_relacion": sum(1 for row in rows if row.clasificacion == "SIN_RELACION"),
                    "bloqueado": sum(1 for row in rows if row.clasificacion == "BLOQUEADO_POR_AMBIGUEDAD"),
                }
            )
        )

    def _write_csv(self, path: Path, rows) -> None:
        fieldnames = [
            "sku_actual",
            "producto_actual",
            "clasificacion",
            "sku_base",
            "producto_base",
            "sku_historico",
            "producto_historico",
            "complemento",
            "regla_costeo",
            "regla_forecast",
            "regla_insumos",
            "confianza",
            "estado",
            "nota_negocio",
            "origen",
        ]
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                writer.writerow(asdict(row))

    def _write_workbook(self, path: Path, rows, generated_at: str) -> None:
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        ws_data = wb.create_sheet("Matriz")

        header_fill = PatternFill("solid", fgColor="6E1F46")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D4B7C5")

        ws_summary["A1"] = "Matriz maestra de relaciones comerciales Point-ERP"
        ws_summary["A1"].font = Font(size=16, bold=True, color="6E1F46")
        ws_summary["A2"] = "Generado"
        ws_summary["B2"] = generated_at
        ws_summary["A4"] = "Clasificacion"
        ws_summary["B4"] = "Cantidad"
        summary_rows = [
            ("HISTORICO_LEGADO", sum(1 for row in rows if row.clasificacion == "HISTORICO_LEGADO")),
            ("COMPLEMENTO_OBLIGATORIO", sum(1 for row in rows if row.clasificacion == "COMPLEMENTO_OBLIGATORIO")),
            ("PRODUCTO_BASE_DIRECTO", sum(1 for row in rows if row.clasificacion == "PRODUCTO_BASE_DIRECTO")),
            ("SIN_RELACION", sum(1 for row in rows if row.clasificacion == "SIN_RELACION")),
            ("BLOQUEADO_POR_AMBIGUEDAD", sum(1 for row in rows if row.clasificacion == "BLOQUEADO_POR_AMBIGUEDAD")),
        ]
        for idx, (label, value) in enumerate(summary_rows, start=5):
            ws_summary[f"A{idx}"] = label
            ws_summary[f"B{idx}"] = value
        ws_summary.column_dimensions["A"].width = 34
        ws_summary.column_dimensions["B"].width = 16

        headers = [
            "sku_actual",
            "producto_actual",
            "clasificacion",
            "sku_base",
            "producto_base",
            "sku_historico",
            "producto_historico",
            "complemento",
            "regla_costeo",
            "regla_forecast",
            "regla_insumos",
            "confianza",
            "estado",
            "nota_negocio",
            "origen",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row_idx, row in enumerate(rows, start=2):
            payload = asdict(row)
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_data.cell(row=row_idx, column=col_idx, value=payload.get(header, ""))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(left=thin, right=thin, bottom=thin)
        for idx, header in enumerate(headers, start=1):
            max_len = max(len(header), *(len(str(getattr(row, header, ""))) for row in rows)) if rows else len(header)
            ws_data.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 16), 42)
        end_col = get_column_letter(len(headers))
        end_row = len(rows) + 1
        table = Table(displayName="tbl_relaciones_comerciales", ref=f"A1:{end_col}{end_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws_data.add_table(table)
        ws_data.freeze_panes = "A2"

        wb.save(path)
