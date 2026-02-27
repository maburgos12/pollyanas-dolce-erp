from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from core.models import Sucursal
from recetas.models import PoliticaStockSucursalProducto, Receta
from recetas.utils.normalizacion import normalizar_nombre


def _clean(value: Any) -> str:
    if value is None:
        return ""
    txt = str(value).strip()
    if txt.lower() in {"none", "nan"}:
        return ""
    return txt


def _to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    txt = _clean(value).replace(",", ".")
    if not txt:
        return None
    try:
        return Decimal(txt)
    except InvalidOperation:
        return None


def _quantize_qty(value: Decimal | None) -> Decimal:
    return max(Decimal("0"), Decimal(str(value or 0))).quantize(Decimal("0.001"))


@dataclass
class Counters:
    read: int = 0
    valid: int = 0
    created: int = 0
    updated: int = 0
    unchanged: int = 0
    skipped: int = 0
    unresolved_sucursal: int = 0
    unresolved_receta: int = 0
    no_stock: int = 0


class Command(BaseCommand):
    help = "Importa políticas de stock mínimo por sucursal-producto desde XLSX normalizado."

    def add_arguments(self, parser):
        parser.add_argument("archivo", type=str, help="Ruta XLSX generado por extraer_stock_minimos_sucursales.")
        parser.add_argument(
            "--sheet",
            type=str,
            default="stock_minimo_abastecimiento",
            help="Nombre de hoja a leer (default: stock_minimo_abastecimiento).",
        )
        parser.add_argument(
            "--strategy",
            choices=["max", "lv", "sd"],
            default="max",
            help="Cómo consolidar LV/SD cuando no hay política por periodo en DB.",
        )
        parser.add_argument("--dry-run", action="store_true", help="Solo valida y muestra conteos sin guardar.")

    def handle(self, *args, **options):
        input_path = Path(options["archivo"]).expanduser()
        if not input_path.exists():
            raise CommandError(f"No existe archivo: {input_path}")
        if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise CommandError("Formato no soportado. Usa XLSX/XLSM.")

        sheet_name = (options.get("sheet") or "stock_minimo_abastecimiento").strip()
        strategy = (options.get("strategy") or "max").strip().lower()
        dry_run = bool(options.get("dry_run"))

        wb = load_workbook(input_path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"No existe hoja '{sheet_name}' en {input_path.name}.")
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("Hoja vacía.")
        headers = [normalizar_nombre(str(c or "")).replace("_", " ") for c in rows[0]]
        idx = {h: i for i, h in enumerate(headers) if h}

        required_any = {"sucursal", "periodo", "stock minimo"}
        if not required_any.issubset(set(idx)):
            raise CommandError(
                "La hoja no tiene columnas requeridas. Se esperaban al menos: sucursal, periodo, stock minimo."
            )

        receta_col = None
        for candidate in ("receta match", "receta"):
            if candidate in idx:
                receta_col = candidate
                break
        if receta_col is None:
            raise CommandError("La hoja no contiene columna de receta (receta match / receta).")

        match_status_col = "match status" if "match status" in idx else None
        codigo_col = "codigo point" if "codigo point" in idx else None

        sucursales = list(Sucursal.objects.filter(activa=True).only("id", "codigo", "nombre"))
        suc_map: dict[str, Sucursal] = {}
        for s in sucursales:
            suc_map[normalizar_nombre(s.codigo)] = s
            suc_map[normalizar_nombre(s.nombre)] = s

        recetas_qs = Receta.objects.filter(tipo=Receta.TIPO_PRODUCTO_FINAL).only(
            "id", "nombre", "nombre_normalizado", "codigo_point"
        )
        receta_by_name: dict[str, Receta] = {}
        receta_by_code: dict[str, Receta] = {}
        for r in recetas_qs:
            receta_by_name[r.nombre_normalizado or normalizar_nombre(r.nombre)] = r
            if r.codigo_point:
                receta_by_code[normalizar_nombre(r.codigo_point)] = r

        counters = Counters()
        unresolved_suc: dict[str, int] = {}
        unresolved_rec: dict[str, int] = {}
        by_key: dict[tuple[int, int], dict[str, Decimal]] = {}

        for raw in rows[1:]:
            counters.read += 1

            status_val = _clean(raw[idx[match_status_col]]) if match_status_col else ""
            if status_val and status_val.upper() != "MATCH_OK":
                counters.skipped += 1
                continue

            suc_txt = _clean(raw[idx["sucursal"]])
            per_txt = _clean(raw[idx["periodo"]]).upper()
            stock = _to_decimal(raw[idx["stock minimo"]])
            receta_txt = _clean(raw[idx[receta_col]])
            codigo_txt = _clean(raw[idx[codigo_col]]) if codigo_col else ""

            if per_txt not in {"LV", "SD"}:
                counters.skipped += 1
                continue
            if stock is None or stock <= 0:
                counters.no_stock += 1
                counters.skipped += 1
                continue
            if not suc_txt or not receta_txt:
                counters.skipped += 1
                continue

            suc = suc_map.get(normalizar_nombre(suc_txt))
            if not suc:
                counters.unresolved_sucursal += 1
                unresolved_suc[suc_txt] = unresolved_suc.get(suc_txt, 0) + 1
                counters.skipped += 1
                continue

            receta = None
            if codigo_txt:
                receta = receta_by_code.get(normalizar_nombre(codigo_txt))
            if receta is None:
                receta = receta_by_name.get(normalizar_nombre(receta_txt))
            if receta is None:
                counters.unresolved_receta += 1
                unresolved_rec[receta_txt] = unresolved_rec.get(receta_txt, 0) + 1
                counters.skipped += 1
                continue

            counters.valid += 1
            key = (suc.id, receta.id)
            if key not in by_key:
                by_key[key] = {"LV": Decimal("0"), "SD": Decimal("0")}
            if stock > by_key[key][per_txt]:
                by_key[key][per_txt] = stock

        @transaction.atomic
        def _persist() -> None:
            for (sucursal_id, receta_id), period_values in by_key.items():
                lv = _quantize_qty(period_values.get("LV"))
                sd = _quantize_qty(period_values.get("SD"))
                if strategy == "lv":
                    base = lv if lv > 0 else sd
                elif strategy == "sd":
                    base = sd if sd > 0 else lv
                else:
                    base = max(lv, sd)
                base = _quantize_qty(base)
                if base <= 0:
                    continue

                policy, created = PoliticaStockSucursalProducto.objects.get_or_create(
                    sucursal_id=sucursal_id,
                    receta_id=receta_id,
                    defaults={
                        "stock_minimo": base,
                        "stock_objetivo": base,
                        "stock_maximo": base,
                        "dias_cobertura": 1,
                        "stock_seguridad": Decimal("0"),
                        "lote_minimo": Decimal("0"),
                        "multiplo_empaque": Decimal("1"),
                        "activa": True,
                    },
                )
                if created:
                    counters.created += 1
                    continue

                changed = False
                if _quantize_qty(policy.stock_minimo) != base:
                    policy.stock_minimo = base
                    changed = True
                if _quantize_qty(policy.stock_objetivo) < base:
                    policy.stock_objetivo = base
                    changed = True
                if _quantize_qty(policy.stock_maximo) < base:
                    policy.stock_maximo = base
                    changed = True
                if not policy.activa:
                    policy.activa = True
                    changed = True

                if changed:
                    policy.save()
                    counters.updated += 1
                else:
                    counters.unchanged += 1

        if not dry_run:
            _persist()

        self.stdout.write(self.style.SUCCESS("Importación de políticas de stock mínimo completada"))
        self.stdout.write(f"  - archivo: {input_path}")
        self.stdout.write(f"  - hoja: {sheet_name}")
        self.stdout.write(f"  - estrategia LV/SD: {strategy}")
        self.stdout.write(f"  - filas leídas: {counters.read}")
        self.stdout.write(f"  - filas válidas: {counters.valid}")
        self.stdout.write(f"  - omitidas: {counters.skipped}")
        self.stdout.write(f"  - creadas: {counters.created}")
        self.stdout.write(f"  - actualizadas: {counters.updated}")
        self.stdout.write(f"  - sin cambios: {counters.unchanged}")
        self.stdout.write(f"  - sucursal no resuelta: {counters.unresolved_sucursal}")
        self.stdout.write(f"  - receta no resuelta: {counters.unresolved_receta}")
        self.stdout.write(f"  - sin stock válido: {counters.no_stock}")

        if unresolved_suc:
            self.stdout.write("  - ejemplos sucursal no resuelta:")
            for name, count in sorted(unresolved_suc.items(), key=lambda x: x[1], reverse=True)[:10]:
                self.stdout.write(f"    * {name}: {count}")
        if unresolved_rec:
            self.stdout.write("  - ejemplos receta no resuelta:")
            for name, count in sorted(unresolved_rec.items(), key=lambda x: x[1], reverse=True)[:15]:
                self.stdout.write(f"    * {name}: {count}")

