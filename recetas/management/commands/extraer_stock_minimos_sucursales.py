from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from openpyxl import Workbook, load_workbook
from rapidfuzz import fuzz

from core.models import Sucursal
from recetas.models import Receta
from recetas.utils.normalizacion import normalizar_nombre


PRESENTACION_MAP = {
    "g": "Grande",
    "grande": "Grande",
    "med": "Mediano",
    "m": "Mediano",
    "mediano": "Mediano",
    "ch": "Chico",
    "chico": "Chico",
    "mini": "Mini",
    "r": "Rebanada",
    "reb": "Rebanada",
    "rebanada": "Rebanada",
    "ind": "Individual",
    "individual": "Individual",
    "bollo": "Bollo",
    "bollos": "Bollo",
    "1/2 plancha": "1/2 Plancha",
    "media plancha": "1/2 Plancha",
}

SUCURSAL_ALIASES = {
    "matriz": "Matriz",
    "leyva": "Leyva",
    "payan": "Payán",
    "glorias": "Las Glorias",
    "las glorias": "Las Glorias",
    "colosio": "Colosio",
    "crucero": "Crucero",
    "nio": "Plaza Nío",
    "plaza nio": "Plaza Nío",
    "tunel": "EL TUNEL",
    "el tunel": "EL TUNEL",
}

BLOCK_CONFIGS = [
    # Lunes-Viernes
    ("LV_BLOQUE_1", "LV", 1, 2, 3, 4, False),
    ("LV_BLOQUE_2", "LV", 8, 9, 10, 11, False),
    # Sábado-Domingo
    ("SD_BLOQUE_1", "SD", 15, 16, 17, 18, False),
    # Este bloque viene mezclado (producto/presentación en dos columnas)
    ("SD_BLOQUE_2", "SD", 22, 23, 24, 25, True),
]


@dataclass
class StockRow:
    sucursal: str
    periodo: str
    categoria: str
    producto: str
    presentacion: str
    stock_minimo: Decimal
    sheet: str
    source_block: str
    row_number: int
    receta_match: str = ""
    codigo_point_match: str = ""
    match_score: int = 0
    match_status: str = "NO_MATCH"

    @property
    def nombre_producto(self) -> str:
        base = f"{self.categoria} {self.producto}".strip()
        return re.sub(r"\s+", " ", base).strip()


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    txt = str(value).strip()
    if txt.lower() in {"none", "nan"}:
        return ""
    return txt


def _to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    raw = _clean_text(value)
    if not raw:
        return None
    # Extrae componente numérico principal (soporta 16L, 2kg, 250pz)
    match = re.search(r"-?\d+(?:[.,]\d+)?", raw)
    if not match:
        return None
    num = match.group(0).replace(",", ".")
    try:
        return Decimal(num)
    except InvalidOperation:
        return None


def _looks_like_presentacion(value: str) -> bool:
    norm = normalizar_nombre(value).replace("_", " ").strip()
    return norm in PRESENTACION_MAP


def _canonical_presentacion(value: str) -> str:
    norm = normalizar_nombre(value).replace("_", " ").strip()
    return PRESENTACION_MAP.get(norm, value.strip())


def _is_noise_name(value: str) -> bool:
    n = normalizar_nombre(value).replace("_", " ").strip()
    if not n:
        return True
    if n in {"producto", "present", "presentacion", "stock", "existencia", "solicitud sugerida"}:
        return True
    if n in {"recibe", "entrega", "insumos cedis", "insumos almacen"}:
        return True
    return False


def _compose_producto(categoria: str, producto: str) -> tuple[str, str]:
    categoria_clean = _clean_text(categoria)
    producto_clean = _clean_text(producto)
    if categoria_clean and producto_clean:
        if normalizar_nombre(producto_clean).startswith(normalizar_nombre(categoria_clean)):
            return "", producto_clean
        return categoria_clean, producto_clean
    if categoria_clean and not producto_clean:
        return "", categoria_clean
    return "", producto_clean


class Command(BaseCommand):
    help = (
        "Extrae stock mínimo por sucursal desde plantilla legacy de stock y genera "
        "consolidado normalizado + plantilla operativa de captura."
    )

    def add_arguments(self, parser):
        parser.add_argument("archivo", type=str, help="Ruta del archivo XLSX legado de stocks por sucursal.")
        parser.add_argument(
            "--output-dir",
            type=str,
            default="output/spreadsheet",
            help="Carpeta de salida para archivos generados.",
        )
        parser.add_argument(
            "--output-suffix",
            type=str,
            default="",
            help="Sufijo fijo para nombre de archivos (si no se envía usa timestamp).",
        )

    def handle(self, *args, **options):
        input_path = Path(options["archivo"]).expanduser()
        if not input_path.exists():
            raise CommandError(f"No existe archivo: {input_path}")
        if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise CommandError("Formato no soportado. Usa .xlsx/.xlsm para extracción de stock mínimo.")

        out_dir = Path(options.get("output_dir") or "output/spreadsheet")
        out_dir.mkdir(parents=True, exist_ok=True)
        suffix = (options.get("output_suffix") or "").strip() or datetime.now().strftime("%Y%m%d_%H%M%S")
        output_xlsx = out_dir / f"stock_minimo_sucursales_{suffix}.xlsx"
        output_csv = out_dir / f"stock_minimo_sucursales_{suffix}.csv"
        output_template = out_dir / f"plantilla_captura_reabasto_sucursales_{suffix}.xlsx"

        wb = load_workbook(input_path, data_only=True)
        sucursal_map = self._build_sucursal_map()
        receta_index = self._build_receta_index()
        rows = self._extract_rows(wb, sucursal_map)
        if not rows:
            raise CommandError("No se encontraron filas de stock mínimo en hojas 'Stock *'.")

        deduped = self._dedupe_rows(rows)
        self._match_recetas(deduped, receta_index)
        self._write_outputs(deduped, output_xlsx, output_csv, output_template)

        total = len(deduped)
        abastecimiento_rows = [r for r in deduped if not self._is_non_production_row(r)]
        matched = sum(1 for r in deduped if r.match_status != "NO_MATCH")
        no_match = total - matched
        matched_ok = sum(1 for r in deduped if r.match_status == "MATCH_OK")
        matched_review = sum(1 for r in deduped if r.match_status == "MATCH_REVISAR")
        lv = sum(1 for r in deduped if r.periodo == "LV")
        sd = sum(1 for r in deduped if r.periodo == "SD")

        self.stdout.write(self.style.SUCCESS("Extracción de stock mínimo completada"))
        self.stdout.write(f"  - archivo fuente: {input_path}")
        self.stdout.write(f"  - filas normalizadas: {total}")
        self.stdout.write(f"  - filas abastecimiento (sin INSUMOS ALMACEN): {len(abastecimiento_rows)}")
        self.stdout.write(f"  - periodo LV: {lv}")
        self.stdout.write(f"  - periodo SD: {sd}")
        self.stdout.write(f"  - match receta OK: {matched_ok}")
        self.stdout.write(f"  - match receta revisar: {matched_review}")
        self.stdout.write(f"  - pendientes match: {no_match}")
        self.stdout.write(f"  - consolidado XLSX: {output_xlsx}")
        self.stdout.write(f"  - consolidado CSV: {output_csv}")
        self.stdout.write(f"  - plantilla captura tiendas: {output_template}")

    def _build_sucursal_map(self) -> dict[str, str]:
        mapping: dict[str, str] = {}
        for suc in Sucursal.objects.filter(activa=True).only("codigo", "nombre").order_by("id"):
            mapping[normalizar_nombre(suc.codigo)] = suc.nombre
            mapping[normalizar_nombre(suc.nombre)] = suc.nombre
        for alias, canonical in SUCURSAL_ALIASES.items():
            mapping[normalizar_nombre(alias)] = canonical
        return mapping

    def _build_receta_index(self) -> list[tuple[str, str, str]]:
        rows: list[tuple[str, str, str]] = []
        qs = (
            Receta.objects.filter(tipo=Receta.TIPO_PRODUCTO_FINAL)
            .only("nombre", "nombre_normalizado", "codigo_point")
            .order_by("id")
        )
        for receta in qs:
            nombre = receta.nombre or ""
            norm = receta.nombre_normalizado or normalizar_nombre(nombre)
            rows.append((nombre, norm, receta.codigo_point or ""))
        return rows

    def _extract_rows(self, workbook, sucursal_map: dict[str, str]) -> list[StockRow]:
        extracted: list[StockRow] = []
        for sheet_name in workbook.sheetnames:
            if not normalizar_nombre(sheet_name).startswith("stock "):
                continue
            ws = workbook[sheet_name]
            sucursal = self._resolve_sucursal_name(ws, sheet_name, sucursal_map)

            for block_id, periodo, cat_col, prod_col, present_col, stock_col, merge_text in BLOCK_CONFIGS:
                carry_cat = ""
                carry_prod = ""
                for row_idx in range(6, ws.max_row + 1):
                    raw_cat = _clean_text(ws.cell(row_idx, cat_col).value) if cat_col else ""
                    raw_prod = _clean_text(ws.cell(row_idx, prod_col).value)
                    raw_present = _clean_text(ws.cell(row_idx, present_col).value)
                    stock_raw = ws.cell(row_idx, stock_col).value
                    stock = _to_decimal(stock_raw)

                    if raw_cat and not _is_noise_name(raw_cat):
                        carry_cat = raw_cat

                    if raw_prod and not _is_noise_name(raw_prod):
                        if merge_text and raw_present and not _looks_like_presentacion(raw_present):
                            carry_prod = f"{raw_prod} {raw_present}".strip()
                        else:
                            carry_prod = raw_prod

                    if stock is None or stock <= 0:
                        continue

                    categoria = raw_cat if raw_cat and not _is_noise_name(raw_cat) else carry_cat
                    producto = raw_prod if raw_prod and not _is_noise_name(raw_prod) else carry_prod
                    presentacion = raw_present

                    if merge_text and presentacion and not _looks_like_presentacion(presentacion):
                        producto = f"{producto} {presentacion}".strip()
                        presentacion = ""

                    categoria, producto = _compose_producto(categoria, producto)
                    if _is_noise_name(producto):
                        continue

                    extracted.append(
                        StockRow(
                            sucursal=sucursal,
                            periodo=periodo,
                            categoria=categoria,
                            producto=producto,
                            presentacion=_canonical_presentacion(presentacion) if presentacion else "",
                            stock_minimo=stock.quantize(Decimal("0.001")),
                            sheet=sheet_name,
                            source_block=block_id,
                            row_number=row_idx,
                        )
                    )
        return extracted

    def _resolve_sucursal_name(self, ws, sheet_name: str, sucursal_map: dict[str, str]) -> str:
        candidates = [
            _clean_text(ws["B3"].value),
            _clean_text(ws["P3"].value),
            re.sub(r"(?i)^stock\\s+", "", sheet_name).strip(),
        ]
        for raw in candidates:
            norm = normalizar_nombre(raw)
            if norm in sucursal_map:
                return sucursal_map[norm]
        return candidates[-1] or sheet_name

    def _dedupe_rows(self, rows: list[StockRow]) -> list[StockRow]:
        by_key: dict[tuple[str, str, str, str], StockRow] = {}
        for row in rows:
            key = (
                row.sucursal,
                row.periodo,
                normalizar_nombre(row.nombre_producto),
                normalizar_nombre(row.presentacion),
            )
            prev = by_key.get(key)
            if prev is None or row.stock_minimo > prev.stock_minimo:
                by_key[key] = row
        result = list(by_key.values())
        result.sort(
            key=lambda r: (
                r.sucursal.lower(),
                r.periodo,
                normalizar_nombre(r.nombre_producto),
                normalizar_nombre(r.presentacion),
            )
        )
        return result

    def _match_recetas(self, rows: list[StockRow], receta_index: list[tuple[str, str, str]]) -> None:
        if not receta_index:
            return
        for row in rows:
            base = row.nombre_producto.strip()
            pres = row.presentacion.strip()
            candidates = []
            if pres:
                candidates.append(f"{base} - {pres}".strip())
                candidates.append(f"Pastel {base} - {pres}".strip())
                candidates.append(f"{base} {pres}".strip())
            candidates.append(base)
            if not normalizar_nombre(base).startswith("pastel"):
                candidates.append(f"Pastel {base}".strip())

            best_score = -1
            best_name = ""
            best_code = ""
            for cand in candidates:
                cand_norm = normalizar_nombre(cand)
                if not cand_norm:
                    continue
                for receta_name, receta_norm, codigo in receta_index:
                    if cand_norm == receta_norm:
                        best_score = 100
                        best_name = receta_name
                        best_code = codigo
                        break
                    # Contención para nombres cortos
                    if cand_norm in receta_norm or receta_norm in cand_norm:
                        score = 95
                    else:
                        score = int(fuzz.token_set_ratio(cand_norm, receta_norm))
                    if score > best_score:
                        best_score = score
                        best_name = receta_name
                        best_code = codigo
                if best_score == 100:
                    break

            row.match_score = max(0, best_score)
            if row.match_score >= 88:
                row.match_status = "MATCH_OK"
                row.receta_match = best_name
                row.codigo_point_match = best_code
            elif row.match_score >= 80:
                row.match_status = "MATCH_REVISAR"
                row.receta_match = best_name
                row.codigo_point_match = best_code
            else:
                row.match_status = "NO_MATCH"
                row.receta_match = best_name if row.match_score > 0 else ""
                row.codigo_point_match = best_code if row.match_score > 0 else ""

    def _write_outputs(
        self,
        rows: list[StockRow],
        output_xlsx: Path,
        output_csv: Path,
        output_template: Path,
    ) -> None:
        self._write_csv(rows, output_csv)
        self._write_xlsx(rows, output_xlsx)
        self._write_template(rows, output_template)

    def _write_csv(self, rows: list[StockRow], output_csv: Path) -> None:
        import csv

        with output_csv.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "sucursal",
                    "periodo",
                    "categoria",
                    "producto",
                    "presentacion",
                    "nombre_producto",
                    "stock_minimo",
                    "receta_match",
                    "codigo_point_match",
                    "match_score",
                    "match_status",
                    "sheet",
                    "source_block",
                    "row_number",
                ]
            )
            for row in rows:
                writer.writerow(
                    [
                        row.sucursal,
                        row.periodo,
                        row.categoria,
                        row.producto,
                        row.presentacion,
                        row.nombre_producto,
                        str(row.stock_minimo),
                        row.receta_match,
                        row.codigo_point_match,
                        row.match_score,
                        row.match_status,
                        row.sheet,
                        row.source_block,
                        row.row_number,
                    ]
                )

    def _write_xlsx(self, rows: list[StockRow], output_xlsx: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "stock_minimo_normalizado"
        headers = [
            "Sucursal",
            "Periodo",
            "Categoria",
            "Producto",
            "Presentacion",
            "Nombre producto",
            "Stock minimo",
            "Receta match",
            "Codigo Point",
            "Match score",
            "Match status",
            "Sheet",
            "Bloque",
            "Fila origen",
        ]
        ws.append(headers)
        for row in rows:
            ws.append(
                [
                    row.sucursal,
                    row.periodo,
                    row.categoria,
                    row.producto,
                    row.presentacion,
                    row.nombre_producto,
                    float(row.stock_minimo),
                    row.receta_match,
                    row.codigo_point_match,
                    row.match_score,
                    row.match_status,
                    row.sheet,
                    row.source_block,
                    row.row_number,
                ]
            )
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:N{ws.max_row}"

        ws_abas = wb.create_sheet("stock_minimo_abastecimiento")
        ws_abas.append(headers)
        for row in rows:
            if self._is_non_production_row(row):
                continue
            ws_abas.append(
                [
                    row.sucursal,
                    row.periodo,
                    row.categoria,
                    row.producto,
                    row.presentacion,
                    row.nombre_producto,
                    float(row.stock_minimo),
                    row.receta_match,
                    row.codigo_point_match,
                    row.match_score,
                    row.match_status,
                    row.sheet,
                    row.source_block,
                    row.row_number,
                ]
            )
        ws_abas.freeze_panes = "A2"
        ws_abas.auto_filter.ref = f"A1:N{max(2, ws_abas.max_row)}"

        ws_recetas = wb.create_sheet("stock_minimo_recetas")
        ws_recetas.append(headers)
        for row in rows:
            if self._is_non_production_row(row):
                continue
            if row.match_status == "NO_MATCH":
                continue
            ws_recetas.append(
                [
                    row.sucursal,
                    row.periodo,
                    row.categoria,
                    row.producto,
                    row.presentacion,
                    row.nombre_producto,
                    float(row.stock_minimo),
                    row.receta_match,
                    row.codigo_point_match,
                    row.match_score,
                    row.match_status,
                    row.sheet,
                    row.source_block,
                    row.row_number,
                ]
            )
        ws_recetas.freeze_panes = "A2"
        ws_recetas.auto_filter.ref = f"A1:N{max(2, ws_recetas.max_row)}"

        self._append_pivot_sheet(wb, rows, "pivot_lv", "LV")
        self._append_pivot_sheet(wb, rows, "pivot_sd", "SD")

        ws_match = wb.create_sheet("pendientes_match")
        ws_match.append(headers)
        for row in rows:
            if row.match_status == "MATCH_OK":
                continue
            ws_match.append(
                [
                    row.sucursal,
                    row.periodo,
                    row.categoria,
                    row.producto,
                    row.presentacion,
                    row.nombre_producto,
                    float(row.stock_minimo),
                    row.receta_match,
                    row.codigo_point_match,
                    row.match_score,
                    row.match_status,
                    row.sheet,
                    row.source_block,
                    row.row_number,
                ]
            )
        ws_match.freeze_panes = "A2"
        ws_match.auto_filter.ref = f"A1:N{max(2, ws_match.max_row)}"

        wb.save(output_xlsx)

    def _append_pivot_sheet(self, wb: Workbook, rows: list[StockRow], sheet_name: str, periodo: str) -> None:
        subset = [r for r in rows if r.periodo == periodo]
        sucursales = sorted({r.sucursal for r in subset})
        keys = sorted(
            {(r.nombre_producto, r.presentacion) for r in subset},
            key=lambda item: (normalizar_nombre(item[0]), normalizar_nombre(item[1])),
        )
        matrix: dict[tuple[str, str], dict[str, Decimal]] = {}
        for r in subset:
            key = (r.nombre_producto, r.presentacion)
            matrix.setdefault(key, {})
            prev = matrix[key].get(r.sucursal, Decimal("0"))
            if r.stock_minimo > prev:
                matrix[key][r.sucursal] = r.stock_minimo

        ws = wb.create_sheet(sheet_name)
        ws.append(["Nombre producto", "Presentacion", *sucursales])
        for key in keys:
            row = [key[0], key[1]]
            for suc in sucursales:
                val = matrix.get(key, {}).get(suc)
                row.append(float(val) if val is not None else "")
            ws.append(row)
        ws.freeze_panes = "A2"
        last_col = 2 + len(sucursales)
        col_letter = ws.cell(1, last_col).column_letter
        ws.auto_filter.ref = f"A1:{col_letter}{max(2, ws.max_row)}"

    def _write_template(self, rows: list[StockRow], output_template: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "captura_reabasto"
        headers = [
            "fecha_operacion",
            "sucursal",
            "periodo",
            "categoria",
            "producto",
            "presentacion",
            "receta",
            "codigo_point",
            "match_status",
            "match_score",
            "stock_minimo",
            "stock_reportado",
            "en_transito",
            "consumo_proyectado",
            "solicitado",
            "justificacion",
            "observaciones",
        ]
        ws.append(headers)

        ok_map: dict[tuple[str, str, str], StockRow] = {}
        review_rows: list[StockRow] = []
        for row in rows:
            if self._is_non_production_row(row):
                continue
            if row.match_status == "MATCH_OK" and row.receta_match:
                key = (row.sucursal, row.periodo, row.receta_match)
                prev = ok_map.get(key)
                if prev is None or row.stock_minimo > prev.stock_minimo:
                    ok_map[key] = row
            elif row.match_status == "MATCH_REVISAR":
                review_rows.append(row)

        rows_for_template = list(ok_map.values()) + review_rows
        rows_for_template.sort(
            key=lambda r: (
                r.sucursal.lower(),
                r.periodo,
                normalizar_nombre(r.receta_match or r.nombre_producto),
                normalizar_nombre(r.presentacion),
            )
        )

        for row in rows_for_template:
            receta_val = row.receta_match if row.match_status == "MATCH_OK" else ""
            code_val = row.codigo_point_match if row.match_status == "MATCH_OK" else ""
            ws.append(
                [
                    "",
                    row.sucursal,
                    row.periodo,
                    row.categoria,
                    row.producto,
                    row.presentacion,
                    receta_val,
                    code_val,
                    row.match_status,
                    row.match_score,
                    float(row.stock_minimo),
                    "",
                    0,
                    0,
                    "",
                    "",
                    "",
                ]
            )
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:Q{max(2, ws.max_row)}"
        wb.save(output_template)

    def _is_non_production_row(self, row: StockRow) -> bool:
        categoria_norm = normalizar_nombre(row.categoria).replace("_", " ")
        producto_norm = normalizar_nombre(row.producto).replace("_", " ")
        nombre_norm = normalizar_nombre(row.nombre_producto).replace("_", " ")

        if "insumos" in categoria_norm:
            return True
        if "almacen" in producto_norm or "almacen" in nombre_norm:
            return True
        return False
