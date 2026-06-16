from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook
from recetas.utils.normalizacion import normalizar_nombre

from rrhh.models import Empleado, MovimientoVacaciones
from rrhh.services_vacaciones import saldo_vacaciones_empleado


@dataclass
class ImportRow:
    empleado_nombre: str
    periodo_anio: int
    periodo_goce_anio: int
    saldo_ciclo: Decimal
    goce_anterior: Decimal
    confirmado: bool
    tratamiento: str


def decimal_value(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0.00")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise CommandError(f"Valor numérico inválido: {value!r}") from exc


def confirmed(value) -> bool:
    return str(value or "").strip().lower().replace("í", "i") == "si"


def year_from_iso(value, default: int) -> int:
    if not value:
        return default
    return int(str(value)[:4])


def read_rows(path: Path, periodo_anio: int) -> list[ImportRow]:
    wb = load_workbook(path, data_only=True)
    if "Saldo inicial" not in wb.sheetnames:
        raise CommandError("El archivo no tiene hoja 'Saldo inicial'.")
    ws = wb["Saldo inicial"]
    headers = {cell.value: idx for idx, cell in enumerate(ws[1])}
    required = {
        "Empleado",
        "Último aniversario cumplido",
        "Días saldo ciclo ERP",
        "Días periodo anterior/vencido",
        "¿Saldo confirmado por RRHH?",
        "Tratamiento para ERP",
    }
    missing = sorted(required - set(headers))
    if missing:
        raise CommandError(f"Faltan columnas: {', '.join(missing)}")

    rows: list[ImportRow] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        empleado_nombre = str(raw[headers["Empleado"]] or "").strip()
        if not empleado_nombre:
            continue
        rows.append(
            ImportRow(
                empleado_nombre=empleado_nombre,
                periodo_anio=periodo_anio,
                periodo_goce_anio=year_from_iso(raw[headers["Último aniversario cumplido"]], periodo_anio - 1),
                saldo_ciclo=decimal_value(raw[headers["Días saldo ciclo ERP"]]),
                goce_anterior=decimal_value(raw[headers["Días periodo anterior/vencido"]]),
                confirmado=confirmed(raw[headers["¿Saldo confirmado por RRHH?"]]),
                tratamiento=str(raw[headers["Tratamiento para ERP"]] or "").strip(),
            )
        )
    return rows


def empleado_por_nombre(nombre: str) -> Empleado:
    normalizado = normalizar_nombre(nombre)
    qs = Empleado.objects.filter(nombre_normalizado=normalizado)
    count = qs.count()
    if count == 1:
        return qs.get()
    if count > 1:
        raise CommandError(f"Empleado duplicado en ERP: {nombre}")
    raise CommandError(f"Empleado no encontrado en ERP: {nombre}")


class Command(BaseCommand):
    help = "Importa saldo inicial de vacaciones desde el Excel conciliado por RRHH."

    def add_arguments(self, parser):
        parser.add_argument("--archivo", required=True)
        parser.add_argument("--periodo", type=int, default=2026)
        parser.add_argument("--import-id", default="saldo-inicial-vacaciones-20260616")
        parser.add_argument("--ejecutar", action="store_true")

    def handle(self, *args, **options):
        path = Path(options["archivo"])
        if not path.exists():
            raise CommandError(f"No existe archivo: {path}")

        rows = read_rows(path, options["periodo"])
        resumen = {"leidas": len(rows), "omitidas": 0, "errores": 0, "ajustes_saldo": 0, "ajustes_goce": 0, "duplicadas": 0}
        import_id = options["import_id"]

        with transaction.atomic():
            for row in rows:
                if not row.confirmado:
                    resumen["omitidas"] += 1
                    self.stdout.write(f"[OMITIR] {row.empleado_nombre}: saldo no confirmado")
                    continue

                try:
                    empleado = empleado_por_nombre(row.empleado_nombre)
                except CommandError as exc:
                    resumen["errores"] += 1
                    self.stdout.write(self.style.ERROR(f"[ERROR] {exc}"))
                    continue

                disponible = saldo_vacaciones_empleado(empleado, periodo_anio=row.periodo_anio)["disponible"]
                ajuste_saldo = row.saldo_ciclo - disponible

                if ajuste_saldo:
                    desc = f"[{import_id}] saldo ciclo ERP {row.periodo_anio}: disponible objetivo {row.saldo_ciclo}"
                    if self._exists(empleado, row.periodo_anio, desc):
                        resumen["duplicadas"] += 1
                    else:
                        resumen["ajustes_saldo"] += 1
                        self._create(empleado, row.periodo_anio, ajuste_saldo, desc, options["ejecutar"])

                if row.goce_anterior:
                    desc = f"[{import_id}] pendiente de goce {row.periodo_goce_anio}: {row.tratamiento}"
                    if self._exists(empleado, row.periodo_goce_anio, desc):
                        resumen["duplicadas"] += 1
                    else:
                        resumen["ajustes_goce"] += 1
                        self._create(empleado, row.periodo_goce_anio, row.goce_anterior, desc[:220], options["ejecutar"])

            if not options["ejecutar"]:
                transaction.set_rollback(True)
            elif resumen["errores"]:
                transaction.set_rollback(True)
                raise CommandError(f"No se aplicó la carga por {resumen['errores']} errores.")

        modo = "APLICADO" if options["ejecutar"] else "DRY-RUN"
        self.stdout.write(self.style.SUCCESS(f"{modo}: {resumen}"))

    def _exists(self, empleado: Empleado, periodo_anio: int, descripcion: str) -> bool:
        return MovimientoVacaciones.objects.filter(
            empleado=empleado,
            periodo_anio=periodo_anio,
            tipo=MovimientoVacaciones.TIPO_AJUSTE,
            descripcion=descripcion[:220],
        ).exists()

    def _create(self, empleado: Empleado, periodo_anio: int, dias: Decimal, descripcion: str, ejecutar: bool):
        self.stdout.write(f"[{'OK' if ejecutar else 'DRY'}] {empleado.nombre} · {periodo_anio} · ajuste {dias} · {descripcion[:90]}")
        if ejecutar:
            MovimientoVacaciones.objects.create(
                empleado=empleado,
                periodo_anio=periodo_anio,
                tipo=MovimientoVacaciones.TIPO_AJUSTE,
                dias=dias,
                descripcion=descripcion[:220],
            )
