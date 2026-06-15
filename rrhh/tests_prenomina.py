from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from core.models import UserModuleAccess
from rrhh.exporters.contpaqi_prenomina import (
    MOVIMIENTOS_HEADERS,
    build_movimientos_contpaqi_rows,
    export_movimientos_contpaqi_xlsx,
    export_revision_xlsx,
)
from rrhh.models import (
    AjusteAsistencia,
    AsistenciaEmpleado,
    Empleado,
    HoraExtra,
    IncidenciaAsistencia,
    PrenominaCorte,
    PrenominaEmpleadoResumen,
    PrenominaEquivalenciaCONTPAQi,
    PrenominaMovimiento,
)
from rrhh.services_ajustes_asistencia import (
    aprobar_ajuste_asistencia,
    crear_ajuste_asistencia,
    rechazar_ajuste_asistencia,
)
from rrhh.services_prenomina import crear_corte_prenomina, recalcular_corte_prenomina


def aware(fecha: date, hora: time):
    return timezone.make_aware(datetime.combine(fecha, hora), timezone.get_current_timezone())


class PrenominaModelTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="paula-prenomina")
        self.empleado = Empleado.objects.create(
            codigo="347",
            nombre="ANAYA BERNAL CARLOS EZEQUIEL",
            fecha_ingreso=date(2026, 6, 10),
            activo=True,
            sucursal="Matriz",
        )
        self.otro_empleado = Empleado.objects.create(
            codigo="348",
            nombre="BARRAZA LOPEZ MARIA",
            fecha_ingreso=date(2026, 6, 10),
            activo=True,
            sucursal="Matriz",
        )

    def test_corte_genera_folio_y_resumen_por_empleado(self):
        corte = PrenominaCorte.objects.create(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            tipo_periodo=PrenominaCorte.TIPO_QUINCENAL,
            creado_por=self.user,
        )
        resumen = PrenominaEmpleadoResumen.objects.create(
            corte=corte,
            empleado=self.empleado,
            dias_periodo=15,
            dias_laborables=6,
            dias_no_laborados_pre_ingreso=9,
            estado=PrenominaEmpleadoResumen.ESTADO_LISTO,
        )

        self.assertTrue(corte.folio.startswith("PRE-202606-"))
        self.assertEqual(str(resumen), f"{corte.folio} · {self.empleado.nombre}")

    def test_corte_genera_folio_secuencial_por_sufijo(self):
        PrenominaCorte.objects.create(
            folio="PRE-202606-009",
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
        )
        corte = PrenominaCorte.objects.create(
            fecha_inicio=date(2026, 6, 16),
            fecha_fin=date(2026, 6, 30),
            fecha_corte=date(2026, 6, 30),
            creado_por=self.user,
        )

        self.assertEqual(corte.folio, "PRE-202606-010")

    def test_movimiento_requiere_equivalencia_para_exportar(self):
        corte = PrenominaCorte.objects.create(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
        )
        movimiento = PrenominaMovimiento.objects.create(
            corte=corte,
            empleado=self.empleado,
            fecha=date(2026, 6, 11),
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_FALTA,
            valor=Decimal("1"),
        )

        self.assertEqual(movimiento.estado, PrenominaMovimiento.ESTADO_PENDIENTE_CONFIGURACION)

        PrenominaEquivalenciaCONTPAQi.objects.create(
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_FALTA,
            clave_contpaqi="F",
            descripcion="Falta",
            aplica_valor=True,
            activo=True,
        )
        movimiento.aplicar_equivalencia()
        movimiento.save(update_fields=["clave_contpaqi", "estado"])

        self.assertEqual(movimiento.clave_contpaqi, "F")
        self.assertEqual(movimiento.estado, PrenominaMovimiento.ESTADO_LISTO)

    def test_movimiento_no_duplica_misma_fuente_tipo_y_corte(self):
        corte = PrenominaCorte.objects.create(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
        )
        base = {
            "corte": corte,
            "empleado": self.empleado,
            "fecha": date(2026, 6, 11),
            "tipo_movimiento_erp": PrenominaMovimiento.TIPO_FALTA,
            "valor": Decimal("1"),
            "fuente_modelo": "IncidenciaAsistencia",
            "fuente_id": "123",
        }
        PrenominaMovimiento.objects.create(**base)

        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                PrenominaMovimiento.objects.create(**base)

    def test_ajuste_asistencia_guarda_valores_anteriores_y_propuestos(self):
        asistencia = AsistenciaEmpleado.objects.create(
            empleado=self.empleado,
            fecha=date(2026, 6, 11),
        )
        ajuste = AjusteAsistencia.objects.create(
            empleado=self.empleado,
            fecha=date(2026, 6, 11),
            asistencia=asistencia,
            tipo_ajuste=AjusteAsistencia.TIPO_SALIDA,
            estado=AjusteAsistencia.ESTADO_PENDIENTE,
            valores_anteriores={"salida": None},
            valores_propuestos={"salida": "2026-06-11T18:05:00-07:00"},
            motivo="Olvido checar salida.",
            solicitado_por=self.user,
        )

        self.assertEqual(ajuste.estado, AjusteAsistencia.ESTADO_PENDIENTE)
        self.assertEqual(ajuste.valores_anteriores["salida"], None)

    def test_ajuste_asistencia_rechaza_asistencia_de_otro_empleado_o_fecha(self):
        asistencia = AsistenciaEmpleado.objects.create(
            empleado=self.otro_empleado,
            fecha=date(2026, 6, 11),
        )

        with self.assertRaises(ValidationError):
            AjusteAsistencia.objects.create(
                empleado=self.empleado,
                fecha=date(2026, 6, 11),
                asistencia=asistencia,
                tipo_ajuste=AjusteAsistencia.TIPO_SALIDA,
                valores_anteriores={"salida": None},
                valores_propuestos={"salida": "2026-06-11T18:05:00-07:00"},
                motivo="Asistencia de otro empleado.",
                solicitado_por=self.user,
            )

        asistencia = AsistenciaEmpleado.objects.create(
            empleado=self.empleado,
            fecha=date(2026, 6, 12),
        )
        with self.assertRaises(ValidationError):
            AjusteAsistencia.objects.create(
                empleado=self.empleado,
                fecha=date(2026, 6, 11),
                asistencia=asistencia,
                tipo_ajuste=AjusteAsistencia.TIPO_SALIDA,
                valores_anteriores={"salida": None},
                valores_propuestos={"salida": "2026-06-11T18:05:00-07:00"},
                motivo="Fecha distinta.",
                solicitado_por=self.user,
            )


class AjusteAsistenciaServiceTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="paula-prenomina")
        self.autorizador = User.objects.create_user(username="rrhh-autorizador")
        self.fecha = date(2026, 6, 11)
        self.empleado = Empleado.objects.create(
            codigo="349",
            nombre="CASTRO LOPEZ ANA",
            fecha_ingreso=date(2026, 6, 1),
            activo=True,
            sucursal="Matriz",
        )

    @patch("rrhh.services_ajustes_asistencia.evaluar_dia_empleado")
    def test_aprobar_ajuste_de_salida_actualiza_asistencia_y_guarda_historial(self, evaluar_mock):
        salida_original = aware(self.fecha, time(17, 45))
        asistencia = AsistenciaEmpleado.objects.create(
            empleado=self.empleado,
            fecha=self.fecha,
            entrada=aware(self.fecha, time(9, 0)),
            salida=salida_original,
        )
        ajuste = crear_ajuste_asistencia(
            empleado=self.empleado,
            fecha=self.fecha,
            tipo_ajuste=AjusteAsistencia.TIPO_SALIDA,
            valores_propuestos={"salida": "2026-06-11T18:05:00-07:00"},
            motivo="Olvido checar salida correcta.",
            solicitado_por=self.user,
        )

        aprobado = aprobar_ajuste_asistencia(ajuste, self.autorizador, comentario="Validado contra checador.")

        asistencia.refresh_from_db()
        aprobado.refresh_from_db()
        self.assertEqual(ajuste.asistencia_id, asistencia.id)
        self.assertEqual(ajuste.valores_anteriores, {"salida": salida_original.isoformat()})
        self.assertEqual(timezone.localtime(asistencia.salida).time(), time(18, 5))
        self.assertEqual(aprobado.estado, AjusteAsistencia.ESTADO_APLICADO)
        self.assertEqual(aprobado.autorizado_por, self.autorizador)
        self.assertEqual(aprobado.aplicado_por, self.autorizador)
        self.assertIsNotNone(aprobado.autorizado_en)
        self.assertIsNotNone(aprobado.aplicado_en)
        self.assertEqual(aprobado.comentario_autorizacion, "Validado contra checador.")
        self.assertEqual(aprobado.valores_aplicados, {"salida": "2026-06-11T18:05:00-07:00"})
        self.assertNotIn("comentario", aprobado.valores_aplicados)
        evaluar_mock.assert_called_once_with(self.empleado, self.fecha)

    @patch("rrhh.services_ajustes_asistencia.evaluar_dia_empleado")
    def test_rechazar_ajuste_no_modifica_asistencia(self, evaluar_mock):
        salida_original = aware(self.fecha, time(17, 45))
        asistencia = AsistenciaEmpleado.objects.create(
            empleado=self.empleado,
            fecha=self.fecha,
            salida=salida_original,
        )
        ajuste = crear_ajuste_asistencia(
            empleado=self.empleado,
            fecha=self.fecha,
            tipo_ajuste=AjusteAsistencia.TIPO_SALIDA,
            valores_propuestos={"salida": "2026-06-11T18:05:00-07:00"},
            motivo="Solicitud por aclarar.",
            solicitado_por=self.user,
        )

        rechazado = rechazar_ajuste_asistencia(ajuste, self.autorizador, comentario="No procede.")

        asistencia.refresh_from_db()
        rechazado.refresh_from_db()
        self.assertEqual(asistencia.salida, salida_original)
        self.assertEqual(rechazado.estado, AjusteAsistencia.ESTADO_RECHAZADO)
        self.assertEqual(rechazado.autorizado_por, self.autorizador)
        self.assertIsNotNone(rechazado.autorizado_en)
        self.assertEqual(rechazado.comentario_autorizacion, "No procede.")
        self.assertEqual(rechazado.valores_aplicados, {})
        self.assertNotIn("comentario", rechazado.valores_aplicados)
        evaluar_mock.assert_not_called()

    @patch("rrhh.services_ajustes_asistencia.evaluar_dia_empleado")
    def test_no_se_puede_aprobar_dos_veces(self, evaluar_mock):
        AsistenciaEmpleado.objects.create(
            empleado=self.empleado,
            fecha=self.fecha,
        )
        ajuste = crear_ajuste_asistencia(
            empleado=self.empleado,
            fecha=self.fecha,
            tipo_ajuste=AjusteAsistencia.TIPO_ENTRADA,
            valores_propuestos={"entrada": "2026-06-11T09:03:00-07:00"},
            motivo="Entrada capturada manualmente.",
            solicitado_por=self.user,
        )
        aprobar_ajuste_asistencia(ajuste, self.autorizador)

        with self.assertRaises(ValidationError):
            aprobar_ajuste_asistencia(ajuste, self.autorizador)

        self.assertEqual(evaluar_mock.call_count, 1)

    def test_crear_ajuste_requiere_motivo(self):
        with self.assertRaises(ValidationError):
            crear_ajuste_asistencia(
                empleado=self.empleado,
                fecha=self.fecha,
                tipo_ajuste=AjusteAsistencia.TIPO_SALIDA,
                valores_propuestos={"salida": "2026-06-11T18:05:00-07:00"},
                motivo=" ",
                solicitado_por=self.user,
            )

    def test_crear_ajuste_rechaza_fecha_previa_a_ingreso(self):
        empleado = Empleado.objects.create(
            codigo="349-B",
            nombre="CASTRO LOPEZ ANA NUEVA",
            fecha_ingreso=date(2026, 6, 12),
            activo=True,
            sucursal="Matriz",
        )

        with self.assertRaises(ValidationError):
            crear_ajuste_asistencia(
                empleado=empleado,
                fecha=date(2026, 6, 11),
                tipo_ajuste=AjusteAsistencia.TIPO_SALIDA,
                valores_propuestos={"salida": "2026-06-11T18:05:00-07:00"},
                motivo="Antes del ingreso real.",
                solicitado_por=self.user,
            )


class PrenominaServiceTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="paula-corte")
        self.empleado = Empleado.objects.create(
            codigo="350",
            nombre="ANAYA BERNAL CARLOS EZEQUIEL",
            fecha_ingreso=date(2026, 6, 10),
            activo=True,
            sucursal="Matriz",
            area="Produccion",
        )

    def test_crear_corte_no_castiga_dias_pre_ingreso(self):
        IncidenciaAsistencia.objects.create(
            empleado=self.empleado,
            fecha=date(2026, 6, 5),
            tipo=IncidenciaAsistencia.TIPO_FALTA,
            estado=IncidenciaAsistencia.ESTADO_CONCILIADO,
            severidad=IncidenciaAsistencia.SEVERIDAD_ALTA,
            detalle="Falta previa a ingreso.",
        )

        corte = crear_corte_prenomina(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
        )
        resumen = corte.resumenes.get(empleado=self.empleado)

        self.assertEqual(resumen.dias_no_laborados_pre_ingreso, 9)
        self.assertEqual(resumen.dias_laborables, 6)
        self.assertEqual(resumen.faltas, 0)
        self.assertEqual(
            corte.movimientos.filter(
                empleado=self.empleado,
                tipo_movimiento_erp=PrenominaMovimiento.TIPO_FALTA,
            ).count(),
            0,
        )

    def test_falta_conciliada_genera_movimiento_si_tiene_equivalencia(self):
        PrenominaEquivalenciaCONTPAQi.objects.create(
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_FALTA,
            clave_contpaqi="F",
            descripcion="Falta",
            aplica_valor=True,
        )
        incidencia = IncidenciaAsistencia.objects.create(
            empleado=self.empleado,
            fecha=date(2026, 6, 11),
            tipo=IncidenciaAsistencia.TIPO_FALTA,
            estado=IncidenciaAsistencia.ESTADO_CONCILIADO,
            severidad=IncidenciaAsistencia.SEVERIDAD_ALTA,
            detalle="Falta validada.",
        )

        corte = crear_corte_prenomina(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
        )

        resumen = corte.resumenes.get(empleado=self.empleado)
        mov = corte.movimientos.get(empleado=self.empleado, tipo_movimiento_erp=PrenominaMovimiento.TIPO_FALTA)
        self.assertEqual(resumen.faltas, 1)
        self.assertEqual(mov.fuente_modelo, "rrhh.IncidenciaAsistencia")
        self.assertEqual(mov.fuente_id, str(incidencia.id))
        self.assertEqual(mov.valor, Decimal("1.00"))
        self.assertEqual(mov.clave_contpaqi, "F")
        self.assertEqual(mov.estado, PrenominaMovimiento.ESTADO_LISTO)

    def test_hora_extra_autorizada_genera_movimiento_y_recalculo_idempotente(self):
        PrenominaEquivalenciaCONTPAQi.objects.create(
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_HORA_EXTRA,
            clave_contpaqi="HE",
            descripcion="Horas extra",
            aplica_horas=True,
        )
        hora_extra = HoraExtra.objects.create(
            empleado=self.empleado,
            fecha=date(2026, 6, 12),
            horas=Decimal("2.50"),
            estado=HoraExtra.ESTADO_AUTORIZADO,
            notas="Autorizadas por jefe.",
        )
        corte = crear_corte_prenomina(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
        )

        recalcular_corte_prenomina(corte)
        resumen = corte.resumenes.get(empleado=self.empleado)
        mov = corte.movimientos.get(empleado=self.empleado, tipo_movimiento_erp=PrenominaMovimiento.TIPO_HORA_EXTRA)

        self.assertEqual(resumen.horas_extra_autorizadas, Decimal("2.50"))
        self.assertEqual(mov.fuente_modelo, "rrhh.HoraExtra")
        self.assertEqual(mov.fuente_id, str(hora_extra.id))
        self.assertEqual(mov.horas, Decimal("2.50"))
        self.assertEqual(mov.clave_contpaqi, "HE")
        self.assertEqual(mov.estado, PrenominaMovimiento.ESTADO_LISTO)
        self.assertEqual(corte.movimientos.filter(tipo_movimiento_erp=PrenominaMovimiento.TIPO_HORA_EXTRA).count(), 1)

    def test_ajuste_pendiente_marca_resumen_en_revision(self):
        crear_ajuste_asistencia(
            empleado=self.empleado,
            fecha=date(2026, 6, 11),
            tipo_ajuste=AjusteAsistencia.TIPO_SALIDA,
            valores_propuestos={"salida": "2026-06-11T18:05:00-07:00"},
            motivo="Pendiente de validar salida.",
            solicitado_por=self.user,
        )

        corte = crear_corte_prenomina(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
        )
        resumen = corte.resumenes.get(empleado=self.empleado)

        self.assertEqual(resumen.ajustes_pendientes, 1)
        self.assertEqual(resumen.dias_asistencia, 0)
        self.assertEqual(resumen.estado, PrenominaEmpleadoResumen.ESTADO_REVISAR)
        self.assertEqual(corte.estado, PrenominaCorte.ESTADO_EN_REVISION)

    def test_movimiento_sin_equivalencia_queda_pendiente_configuracion(self):
        IncidenciaAsistencia.objects.create(
            empleado=self.empleado,
            fecha=date(2026, 6, 11),
            tipo=IncidenciaAsistencia.TIPO_SUSPENSION,
            estado=IncidenciaAsistencia.ESTADO_CONCILIADO,
            severidad=IncidenciaAsistencia.SEVERIDAD_MEDIA,
            detalle="Suspension conciliada.",
        )

        corte = crear_corte_prenomina(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
        )
        mov = corte.movimientos.get(
            empleado=self.empleado,
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_SUSPENSION,
        )

        self.assertEqual(mov.clave_contpaqi, "")
        self.assertEqual(mov.estado, PrenominaMovimiento.ESTADO_PENDIENTE_CONFIGURACION)
        self.assertEqual(corte.resumen["movimientos_pendientes_configuracion"], 1)
        self.assertEqual(corte.estado, PrenominaCorte.ESTADO_EN_REVISION)

    def test_incidencia_resuelta_no_impacta_prenomina(self):
        IncidenciaAsistencia.objects.create(
            empleado=self.empleado,
            fecha=date(2026, 6, 11),
            tipo=IncidenciaAsistencia.TIPO_FALTA,
            estado=IncidenciaAsistencia.ESTADO_RESUELTO,
            severidad=IncidenciaAsistencia.SEVERIDAD_ALTA,
            detalle="Falta ya resuelta.",
        )

        corte = crear_corte_prenomina(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
        )
        resumen = corte.resumenes.get(empleado=self.empleado)

        self.assertEqual(resumen.faltas, 0)
        self.assertEqual(resumen.alertas_bloqueantes, 0)
        self.assertFalse(corte.movimientos.filter(empleado=self.empleado).exists())

    def test_recalcular_preserva_movimiento_manual_con_importe(self):
        corte = crear_corte_prenomina(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
        )
        manual = PrenominaMovimiento.objects.create(
            corte=corte,
            empleado=self.empleado,
            fecha=date(2026, 6, 14),
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_INCAPACIDAD,
            estado=PrenominaMovimiento.ESTADO_LISTO,
            clave_contpaqi="MANUAL",
            importe=Decimal("325.75"),
            fuente_modelo="",
            fuente_id="",
            referencia="captura-manual",
            notas="Movimiento manual capturado por RRHH.",
            metadata={"origen": "manual"},
        )

        recalcular_corte_prenomina(corte)
        manual.refresh_from_db()

        self.assertEqual(manual.importe, Decimal("325.75"))
        self.assertEqual(manual.clave_contpaqi, "MANUAL")
        self.assertEqual(manual.estado, PrenominaMovimiento.ESTADO_LISTO)
        self.assertEqual(manual.referencia, "captura-manual")
        self.assertEqual(manual.notas, "Movimiento manual capturado por RRHH.")
        self.assertEqual(manual.metadata, {"origen": "manual"})

    def test_recalcular_elimina_automatico_obsoleto_y_preserva_manual(self):
        corte = crear_corte_prenomina(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
        )
        automatico = PrenominaMovimiento.objects.create(
            corte=corte,
            empleado=self.empleado,
            fecha=date(2026, 6, 12),
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_FALTA,
            estado=PrenominaMovimiento.ESTADO_LISTO,
            clave_contpaqi="F",
            valor=Decimal("1"),
            fuente_modelo="rrhh.IncidenciaAsistencia",
            fuente_id="9999",
            referencia="rrhh.IncidenciaAsistencia:9999",
        )
        manual = PrenominaMovimiento.objects.create(
            corte=corte,
            empleado=self.empleado,
            fecha=date(2026, 6, 13),
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_INCAPACIDAD,
            estado=PrenominaMovimiento.ESTADO_LISTO,
            clave_contpaqi="INC",
            importe=Decimal("100.00"),
            referencia="manual-incapacidad",
        )

        recalcular_corte_prenomina(corte)

        self.assertFalse(PrenominaMovimiento.objects.filter(pk=automatico.pk).exists())
        self.assertTrue(PrenominaMovimiento.objects.filter(pk=manual.pk).exists())

    def test_recalcular_no_permite_corte_exportado_o_cerrado(self):
        corte = crear_corte_prenomina(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
        )
        corte.estado = PrenominaCorte.ESTADO_EXPORTADO
        corte.save(update_fields=["estado", "actualizado_en"])

        with self.assertRaises(ValidationError):
            recalcular_corte_prenomina(corte)

        corte.estado = PrenominaCorte.ESTADO_CERRADO
        corte.save(update_fields=["estado", "actualizado_en"])

        with self.assertRaises(ValidationError):
            recalcular_corte_prenomina(corte)

    def test_recalcular_preserva_movimiento_automatico_exportado(self):
        incidencia = IncidenciaAsistencia.objects.create(
            empleado=self.empleado,
            fecha=date(2026, 6, 12),
            tipo=IncidenciaAsistencia.TIPO_FALTA,
            estado=IncidenciaAsistencia.ESTADO_CONCILIADO,
            severidad=IncidenciaAsistencia.SEVERIDAD_ALTA,
            detalle="Falta exportada previamente.",
        )
        corte = crear_corte_prenomina(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
        )
        exportado = corte.movimientos.get(
            fuente_modelo="rrhh.IncidenciaAsistencia",
            fuente_id=str(incidencia.id),
        )
        exportado.estado = PrenominaMovimiento.ESTADO_EXPORTADO
        exportado.clave_contpaqi = "F"
        exportado.save(update_fields=["estado", "clave_contpaqi", "actualizado_en"])

        corte = recalcular_corte_prenomina(corte)
        exportado.refresh_from_db()

        self.assertEqual(exportado.estado, PrenominaMovimiento.ESTADO_EXPORTADO)
        self.assertEqual(exportado.clave_contpaqi, "F")
        self.assertEqual(corte.resumen["movimientos_exportados"], 1)
        self.assertEqual(corte.estado, PrenominaCorte.ESTADO_EN_REVISION)

    def test_movimiento_bloqueado_mantiene_corte_en_revision(self):
        corte = crear_corte_prenomina(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
        )
        PrenominaMovimiento.objects.create(
            corte=corte,
            empleado=self.empleado,
            fecha=date(2026, 6, 13),
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_INCAPACIDAD,
            estado=PrenominaMovimiento.ESTADO_BLOQUEADO,
            clave_contpaqi="INC",
            referencia="manual-bloqueado",
        )

        corte = recalcular_corte_prenomina(corte)

        self.assertEqual(corte.resumen["movimientos_bloqueados"], 1)
        self.assertEqual(corte.estado, PrenominaCorte.ESTADO_EN_REVISION)


class PrenominaExportTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="paula-export")
        self.empleado = Empleado.objects.create(
            codigo="351",
            nombre="ANAYA EXPORT",
            fecha_ingreso=date(2026, 6, 1),
            activo=True,
            sucursal="Matriz",
        )
        self.corte = PrenominaCorte.objects.create(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
            resumen={"movimientos_listos": 1},
        )

    def test_build_movimientos_contpaqi_rows_usa_layout_y_campos_reales(self):
        PrenominaMovimiento.objects.create(
            corte=self.corte,
            empleado=self.empleado,
            fecha=date(2026, 6, 11),
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_FALTA,
            clave_contpaqi="F",
            valor=Decimal("1.00"),
            estado=PrenominaMovimiento.ESTADO_LISTO,
            referencia="rrhh.IncidenciaAsistencia:1",
            notas="Falta validada.",
        )
        PrenominaMovimiento.objects.create(
            corte=self.corte,
            empleado=self.empleado,
            fecha=date(2026, 6, 12),
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_SUSPENSION,
            clave_contpaqi="",
            valor=Decimal("1.00"),
            estado=PrenominaMovimiento.ESTADO_PENDIENTE_CONFIGURACION,
            referencia="rrhh.IncidenciaAsistencia:2",
            notas="Pendiente de clave.",
        )

        rows = build_movimientos_contpaqi_rows(self.corte)

        self.assertEqual(rows[0], MOVIMIENTOS_HEADERS)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1], [
            "351",
            "2026-06-11",
            11,
            PrenominaMovimiento.TIPO_FALTA,
            "F",
            "1.00",
            "",
            "",
            "rrhh.IncidenciaAsistencia:1",
            "Falta validada.",
        ])

    def test_build_movimientos_contpaqi_rows_exporta_nulos_como_vacio(self):
        PrenominaMovimiento.objects.create(
            corte=self.corte,
            empleado=self.empleado,
            fecha=date(2026, 6, 13),
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_HORA_EXTRA,
            clave_contpaqi="HE",
            estado=PrenominaMovimiento.ESTADO_LISTO,
            referencia="",
            notas="",
        )

        rows = build_movimientos_contpaqi_rows(self.corte)

        self.assertEqual(rows[1][5], "")
        self.assertEqual(rows[1][6], "")
        self.assertEqual(rows[1][7], "")
        self.assertEqual(rows[1][8], "")
        self.assertEqual(rows[1][9], "")

    def test_export_movimientos_contpaqi_xlsx_response_y_hoja(self):
        PrenominaMovimiento.objects.create(
            corte=self.corte,
            empleado=self.empleado,
            fecha=date(2026, 6, 11),
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_FALTA,
            clave_contpaqi="F",
            valor=Decimal("1.00"),
            estado=PrenominaMovimiento.ESTADO_LISTO,
            referencia="rrhh.IncidenciaAsistencia:1",
            notas="Falta validada.",
        )

        response = export_movimientos_contpaqi_xlsx(self.corte)
        workbook = load_workbook(BytesIO(response.content))

        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", response["Content-Type"])
        self.assertIn(f"{self.corte.folio}_movimientos_contpaqi.xlsx", response["Content-Disposition"])
        self.assertEqual(workbook.sheetnames, ["Movimientos_CONTPAQi"])
        self.assertEqual([cell.value for cell in workbook["Movimientos_CONTPAQi"][1]], MOVIMIENTOS_HEADERS)

    def test_export_revision_xlsx_incluye_hojas_basicas(self):
        PrenominaEmpleadoResumen.objects.create(
            corte=self.corte,
            empleado=self.empleado,
            dias_periodo=15,
            dias_laborables=15,
            dias_asistencia=1,
            faltas=1,
            estado=PrenominaEmpleadoResumen.ESTADO_LISTO,
        )
        PrenominaMovimiento.objects.create(
            corte=self.corte,
            empleado=self.empleado,
            fecha=date(2026, 6, 11),
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_FALTA,
            clave_contpaqi="F",
            valor=Decimal("1.00"),
            estado=PrenominaMovimiento.ESTADO_LISTO,
            referencia="rrhh.IncidenciaAsistencia:1",
            notas="Falta validada.",
        )

        response = export_revision_xlsx(self.corte)
        workbook = load_workbook(BytesIO(response.content))

        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", response["Content-Type"])
        self.assertIn(f"{self.corte.folio}_revision_prenomina.xlsx", response["Content-Disposition"])
        self.assertIn("Resumen", workbook.sheetnames)
        self.assertIn("Empleados", workbook.sheetnames)
        self.assertIn("Movimientos_CONTPAQi", workbook.sheetnames)
        self.assertEqual(workbook["Resumen"]["A1"].value, "Campo")
        self.assertEqual(workbook["Empleados"]["A1"].value, "Codigo")
        self.assertEqual([cell.value for cell in workbook["Movimientos_CONTPAQi"][1]], MOVIMIENTOS_HEADERS)


class PrenominaViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="paula-view",
            is_superuser=True,
            is_staff=True,
        )
        self.client.force_login(self.user)
        self.empleado = Empleado.objects.create(
            codigo="352",
            nombre="ANAYA VIEW",
            fecha_ingreso=date(2026, 6, 1),
            activo=True,
            sucursal="Matriz",
        )

    def _crear_corte(self, resumen=None):
        corte = PrenominaCorte.objects.create(
            fecha_inicio=date(2026, 6, 1),
            fecha_fin=date(2026, 6, 15),
            fecha_corte=date(2026, 6, 15),
            creado_por=self.user,
            resumen=resumen or {"colaboradores": 1, "movimientos_listos": 1},
        )
        PrenominaEmpleadoResumen.objects.create(
            corte=corte,
            empleado=self.empleado,
            dias_periodo=15,
            dias_laborables=15,
            dias_asistencia=1,
            estado=PrenominaEmpleadoResumen.ESTADO_LISTO,
        )
        return corte

    def test_prenomina_list_renderiza_formulario_y_tab(self):
        response = self.client.get(reverse("rrhh:prenomina"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Generar corte")
        self.assertContains(response, "Prenómina")

    def test_crear_corte_redirige_a_detalle(self):
        response = self.client.post(
            reverse("rrhh:prenomina"),
            {
                "fecha_inicio": "2026-06-01",
                "fecha_fin": "2026-06-15",
                "fecha_corte": "2026-06-15",
                "tipo_periodo": PrenominaCorte.TIPO_QUINCENAL,
            },
        )
        corte = PrenominaCorte.objects.get()

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("rrhh:prenomina_detail", kwargs={"pk": corte.pk}))

    def test_prenomina_detail_renderiza_tablas_y_recalcula(self):
        corte = self._crear_corte()

        response = self.client.get(reverse("rrhh:prenomina_detail", kwargs={"pk": corte.pk}))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mesa de cierre")
        self.assertContains(response, "Colaboradores")

        response = self.client.post(
            reverse("rrhh:prenomina_detail", kwargs={"pk": corte.pk}),
            {"action": "recalcular"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("rrhh:prenomina_detail", kwargs={"pk": corte.pk}))

    def test_recalcular_corte_exportado_redirige_sin_500(self):
        corte = self._crear_corte()
        corte.estado = PrenominaCorte.ESTADO_EXPORTADO
        corte.save(update_fields=["estado", "actualizado_en"])

        response = self.client.post(
            reverse("rrhh:prenomina_detail", kwargs={"pk": corte.pk}),
            {"action": "recalcular"},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("rrhh:prenomina_detail", kwargs={"pk": corte.pk}))

    def test_prenomina_persona_renderiza_version_imprimible(self):
        corte = self._crear_corte()

        response = self.client.get(
            reverse("rrhh:prenomina_persona", kwargs={"pk": corte.pk, "empleado_id": self.empleado.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "REVISIÓN INDIVIDUAL DE PRENÓMINA")
        self.assertContains(response, self.empleado.nombre)
        self.assertContains(response, "Imprimir")

    def test_prenomina_ajuste_crear_post_crea_pendiente(self):
        corte = self._crear_corte()

        response = self.client.post(
            reverse("rrhh:prenomina_ajuste_crear", kwargs={"pk": corte.pk, "empleado_id": self.empleado.pk}),
            {
                "fecha": "2026-06-11",
                "tipo_ajuste": AjusteAsistencia.TIPO_SALIDA,
                "valor_propuesto": "2026-06-11T18:05",
                "motivo": "Salida registrada por RRHH",
            },
        )

        ajuste = AjusteAsistencia.objects.get()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            reverse("rrhh:prenomina_persona", kwargs={"pk": corte.pk, "empleado_id": self.empleado.pk}),
        )
        self.assertEqual(ajuste.estado, AjusteAsistencia.ESTADO_PENDIENTE)
        self.assertEqual(ajuste.tipo_ajuste, AjusteAsistencia.TIPO_SALIDA)
        self.assertEqual(ajuste.valores_propuestos["salida"], "2026-06-11T18:05")

    def test_prenomina_ajuste_crear_rechaza_valor_fuera_del_dia(self):
        corte = self._crear_corte()

        response = self.client.post(
            reverse("rrhh:prenomina_ajuste_crear", kwargs={"pk": corte.pk, "empleado_id": self.empleado.pk}),
            {
                "fecha": "2026-06-11",
                "tipo_ajuste": AjusteAsistencia.TIPO_SALIDA,
                "valor_propuesto": "2026-06-20T18:05",
                "motivo": "Salida con fecha incorrecta",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(AjusteAsistencia.objects.count(), 0)

    def test_prenomina_ajuste_crear_rechaza_fecha_fuera_del_corte_y_tipo_invalido(self):
        corte = self._crear_corte()

        fuera_periodo = self.client.post(
            reverse("rrhh:prenomina_ajuste_crear", kwargs={"pk": corte.pk, "empleado_id": self.empleado.pk}),
            {
                "fecha": "2026-06-20",
                "tipo_ajuste": AjusteAsistencia.TIPO_SALIDA,
                "valor_propuesto": "2026-06-20T18:05",
                "motivo": "Fuera de corte",
            },
        )
        tipo_invalido = self.client.post(
            reverse("rrhh:prenomina_ajuste_crear", kwargs={"pk": corte.pk, "empleado_id": self.empleado.pk}),
            {
                "fecha": "2026-06-11",
                "tipo_ajuste": AjusteAsistencia.TIPO_TURNO,
                "valor_propuesto": "2026-06-11T18:05",
                "motivo": "Tipo no soportado en prenómina",
            },
        )

        self.assertEqual(fuera_periodo.status_code, 302)
        self.assertEqual(tipo_invalido.status_code, 302)
        self.assertEqual(AjusteAsistencia.objects.count(), 0)

    def test_prenomina_ajuste_aprobar_post_aplica_y_actualiza_asistencia(self):
        corte = self._crear_corte()
        fecha = date(2026, 6, 11)
        ajuste = crear_ajuste_asistencia(
            self.empleado,
            fecha,
            AjusteAsistencia.TIPO_SALIDA,
            {"salida": "2026-06-11T18:05"},
            "Corregir salida",
            self.user,
        )

        response = self.client.post(
            reverse("rrhh:prenomina_ajuste_aprobar", kwargs={"pk": corte.pk, "ajuste_id": ajuste.pk}),
            {"comentario": "Validado contra checador"},
        )

        ajuste.refresh_from_db()
        asistencia = AsistenciaEmpleado.objects.get(empleado=self.empleado, fecha=fecha)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url,
            reverse("rrhh:prenomina_persona", kwargs={"pk": corte.pk, "empleado_id": self.empleado.pk}),
        )
        self.assertEqual(ajuste.estado, AjusteAsistencia.ESTADO_APLICADO)
        self.assertEqual(ajuste.comentario_autorizacion, "Validado contra checador")
        self.assertEqual(timezone.localtime(asistencia.salida).time().replace(second=0, microsecond=0), time(18, 5))

    def test_prenomina_ajuste_rechazar_post_no_aplica_asistencia(self):
        corte = self._crear_corte()
        fecha = date(2026, 6, 11)
        salida_original = aware(fecha, time(17, 30))
        asistencia = AsistenciaEmpleado.objects.create(empleado=self.empleado, fecha=fecha, salida=salida_original)
        ajuste = crear_ajuste_asistencia(
            self.empleado,
            fecha,
            AjusteAsistencia.TIPO_SALIDA,
            {"salida": "2026-06-11T18:05"},
            "No coincide con checador",
            self.user,
        )

        response = self.client.post(
            reverse("rrhh:prenomina_ajuste_rechazar", kwargs={"pk": corte.pk, "ajuste_id": ajuste.pk}),
            {"comentario": "Sin evidencia"},
        )

        ajuste.refresh_from_db()
        asistencia.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(ajuste.estado, AjusteAsistencia.ESTADO_RECHAZADO)
        self.assertEqual(ajuste.comentario_autorizacion, "Sin evidencia")
        self.assertEqual(asistencia.salida, salida_original)

    def test_prenomina_ajuste_aprobar_rechazar_fuera_de_corte_devuelve_404(self):
        corte = self._crear_corte()
        ajuste = crear_ajuste_asistencia(
            self.empleado,
            date(2026, 6, 20),
            AjusteAsistencia.TIPO_SALIDA,
            {"salida": "2026-06-20T18:05"},
            "Fuera de corte",
            self.user,
        )

        aprobar_response = self.client.post(
            reverse("rrhh:prenomina_ajuste_aprobar", kwargs={"pk": corte.pk, "ajuste_id": ajuste.pk}),
            {"comentario": "No debe aplicar"},
        )
        rechazar_response = self.client.post(
            reverse("rrhh:prenomina_ajuste_rechazar", kwargs={"pk": corte.pk, "ajuste_id": ajuste.pk}),
            {"comentario": "No debe rechazar"},
        )

        ajuste.refresh_from_db()
        self.assertEqual(aprobar_response.status_code, 404)
        self.assertEqual(rechazar_response.status_code, 404)
        self.assertEqual(ajuste.estado, AjusteAsistencia.ESTADO_PENDIENTE)

    def test_prenomina_ajuste_crear_bloquea_corte_exportado(self):
        corte = self._crear_corte()
        corte.estado = PrenominaCorte.ESTADO_EXPORTADO
        corte.save(update_fields=["estado", "actualizado_en"])

        response = self.client.post(
            reverse("rrhh:prenomina_ajuste_crear", kwargs={"pk": corte.pk, "empleado_id": self.empleado.pk}),
            {
                "fecha": "2026-06-11",
                "tipo_ajuste": AjusteAsistencia.TIPO_SALIDA,
                "valor_propuesto": "2026-06-11T18:05",
                "motivo": "Corte exportado",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(AjusteAsistencia.objects.count(), 0)

    def test_prenomina_ajuste_crear_rechaza_fecha_previa_a_ingreso(self):
        corte = self._crear_corte()
        self.empleado.fecha_ingreso = date(2026, 6, 12)
        self.empleado.save(update_fields=["fecha_ingreso"])

        response = self.client.post(
            reverse("rrhh:prenomina_ajuste_crear", kwargs={"pk": corte.pk, "empleado_id": self.empleado.pk}),
            {
                "fecha": "2026-06-11",
                "tipo_ajuste": AjusteAsistencia.TIPO_SALIDA,
                "valor_propuesto": "2026-06-11T18:05",
                "motivo": "Antes de ingreso",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(AjusteAsistencia.objects.count(), 0)
        self.assertFalse(AsistenciaEmpleado.objects.filter(empleado=self.empleado, fecha=date(2026, 6, 11)).exists())

    def test_prenomina_ajuste_aprobar_rechazar_bloquean_corte_exportado(self):
        corte = self._crear_corte()
        corte.estado = PrenominaCorte.ESTADO_EXPORTADO
        corte.save(update_fields=["estado", "actualizado_en"])
        ajuste_aprobar = crear_ajuste_asistencia(
            self.empleado,
            date(2026, 6, 11),
            AjusteAsistencia.TIPO_SALIDA,
            {"salida": "2026-06-11T18:05"},
            "Aprobar en corte exportado",
            self.user,
        )
        ajuste_rechazar = crear_ajuste_asistencia(
            self.empleado,
            date(2026, 6, 12),
            AjusteAsistencia.TIPO_SALIDA,
            {"salida": "2026-06-12T18:05"},
            "Rechazar en corte exportado",
            self.user,
        )

        aprobar_response = self.client.post(
            reverse("rrhh:prenomina_ajuste_aprobar", kwargs={"pk": corte.pk, "ajuste_id": ajuste_aprobar.pk}),
            {"comentario": "Se aplica aunque el corte no recalcula"},
        )
        rechazar_response = self.client.post(
            reverse("rrhh:prenomina_ajuste_rechazar", kwargs={"pk": corte.pk, "ajuste_id": ajuste_rechazar.pk}),
            {"comentario": "Se rechaza aunque el corte no recalcula"},
        )

        ajuste_aprobar.refresh_from_db()
        ajuste_rechazar.refresh_from_db()
        self.assertEqual(aprobar_response.status_code, 302)
        self.assertEqual(rechazar_response.status_code, 302)
        self.assertEqual(ajuste_aprobar.estado, AjusteAsistencia.ESTADO_PENDIENTE)
        self.assertEqual(ajuste_rechazar.estado, AjusteAsistencia.ESTADO_PENDIENTE)
        self.assertFalse(
            AsistenciaEmpleado.objects.filter(
                empleado=self.empleado,
                fecha=date(2026, 6, 11),
                salida__isnull=False,
            ).exists()
        )

    def test_usuario_solo_lectura_no_puede_crear_aprobar_ni_rechazar_ajustes(self):
        corte = self._crear_corte()
        ajuste = crear_ajuste_asistencia(
            self.empleado,
            date(2026, 6, 11),
            AjusteAsistencia.TIPO_SALIDA,
            {"salida": "2026-06-11T18:05"},
            "Corregir salida",
            self.user,
        )
        viewer = User.objects.create_user(username="paula-ajustes-view-only")
        UserModuleAccess.objects.create(
            user=viewer,
            module="rrhh",
            access=UserModuleAccess.ACCESS_VIEW,
            updated_by=self.user,
        )

        self.client.force_login(viewer)
        crear_response = self.client.post(
            reverse("rrhh:prenomina_ajuste_crear", kwargs={"pk": corte.pk, "empleado_id": self.empleado.pk}),
            {
                "fecha": "2026-06-11",
                "tipo_ajuste": AjusteAsistencia.TIPO_SALIDA,
                "valor_propuesto": "2026-06-11T18:05",
                "motivo": "Intento view only",
            },
        )
        aprobar_response = self.client.post(
            reverse("rrhh:prenomina_ajuste_aprobar", kwargs={"pk": corte.pk, "ajuste_id": ajuste.pk}),
            {"comentario": "No autorizado"},
        )
        rechazar_response = self.client.post(
            reverse("rrhh:prenomina_ajuste_rechazar", kwargs={"pk": corte.pk, "ajuste_id": ajuste.pk}),
            {"comentario": "No autorizado"},
        )

        ajuste.refresh_from_db()
        self.assertEqual(crear_response.status_code, 403)
        self.assertEqual(aprobar_response.status_code, 403)
        self.assertEqual(rechazar_response.status_code, 403)
        self.assertEqual(AjusteAsistencia.objects.count(), 1)
        self.assertEqual(ajuste.estado, AjusteAsistencia.ESTADO_PENDIENTE)

    def test_export_revision_responde_xlsx(self):
        corte = self._crear_corte()

        response = self.client.get(reverse("rrhh:prenomina_export_revision", kwargs={"pk": corte.pk}))

        self.assertEqual(response.status_code, 200)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", response["Content-Type"])
        self.assertIn(f"{corte.folio}_revision_prenomina.xlsx", response["Content-Disposition"])

    def test_export_contpaqi_bloquea_si_hay_pendientes(self):
        corte = self._crear_corte(resumen={"movimientos_pendientes_configuracion": 1})
        PrenominaMovimiento.objects.create(
            corte=corte,
            empleado=self.empleado,
            fecha=date(2026, 6, 11),
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_FALTA,
            estado=PrenominaMovimiento.ESTADO_PENDIENTE_CONFIGURACION,
            valor=Decimal("1.00"),
        )

        response = self.client.get(reverse("rrhh:prenomina_export_contpaqi", kwargs={"pk": corte.pk}))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("rrhh:prenomina_detail", kwargs={"pk": corte.pk}))

    def test_export_contpaqi_no_revienta_con_resumen_no_numerico(self):
        corte = self._crear_corte(resumen={"movimientos_pendientes_configuracion": "N/A"})
        PrenominaMovimiento.objects.create(
            corte=corte,
            empleado=self.empleado,
            fecha=date(2026, 6, 11),
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_FALTA,
            estado=PrenominaMovimiento.ESTADO_PENDIENTE_CONFIGURACION,
            valor=Decimal("1.00"),
        )

        response = self.client.get(reverse("rrhh:prenomina_export_contpaqi", kwargs={"pk": corte.pk}))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("rrhh:prenomina_detail", kwargs={"pk": corte.pk}))

    def test_export_contpaqi_marca_corte_y_movimientos_como_exportados(self):
        corte = self._crear_corte()
        movimiento = PrenominaMovimiento.objects.create(
            corte=corte,
            empleado=self.empleado,
            fecha=date(2026, 6, 11),
            tipo_movimiento_erp=PrenominaMovimiento.TIPO_FALTA,
            clave_contpaqi="F",
            estado=PrenominaMovimiento.ESTADO_LISTO,
            valor=Decimal("1.00"),
        )

        response = self.client.get(reverse("rrhh:prenomina_export_contpaqi", kwargs={"pk": corte.pk}))

        corte.refresh_from_db()
        movimiento.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(corte.estado, PrenominaCorte.ESTADO_EXPORTADO)
        self.assertEqual(corte.resumen["movimientos_listos"], 0)
        self.assertEqual(corte.resumen["movimientos_exportados"], 1)
        self.assertEqual(movimiento.estado, PrenominaMovimiento.ESTADO_EXPORTADO)

    def test_usuario_solo_lectura_no_exporta_contpaqi(self):
        viewer = User.objects.create_user(username="paula-view-only")
        UserModuleAccess.objects.create(
            user=viewer,
            module="rrhh",
            access=UserModuleAccess.ACCESS_VIEW,
            updated_by=self.user,
        )
        corte = self._crear_corte()

        self.client.force_login(viewer)
        response = self.client.get(reverse("rrhh:prenomina_export_contpaqi", kwargs={"pk": corte.pk}))

        self.assertEqual(response.status_code, 403)

    def test_usuario_con_nomina_bloqueada_no_entra_por_url_directa(self):
        user = User.objects.create_user(username="paula-sin-nomina")
        UserModuleAccess.objects.create(
            user=user,
            module="rrhh",
            access=UserModuleAccess.ACCESS_MANAGE,
            updated_by=self.user,
        )
        UserModuleAccess.objects.create(
            user=user,
            module="rrhh.nomina",
            access=UserModuleAccess.ACCESS_NONE,
            updated_by=self.user,
        )

        self.client.force_login(user)
        response = self.client.get(reverse("rrhh:prenomina"))

        self.assertEqual(response.status_code, 403)
