from __future__ import annotations

from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook

from rrhh.models import PrenominaMovimiento


XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MOVIMIENTOS_HEADERS = [
    "CodigoEmpleado",
    "Fecha",
    "Dia",
    "TipoMovimiento",
    "ClaveCONTPAQi",
    "Valor",
    "Horas",
    "Importe",
    "ReferenciaERP",
    "Observaciones",
]


def _blank_if_none(value):
    if value is None:
        return ""
    return str(value)


def build_movimientos_contpaqi_rows(corte):
    rows = [MOVIMIENTOS_HEADERS]
    movimientos = (
        corte.movimientos.select_related("empleado")
        .filter(estado=PrenominaMovimiento.ESTADO_LISTO)
        .order_by("empleado__codigo", "empleado__nombre", "fecha", "id")
    )
    for mov in movimientos:
        rows.append(
            [
                mov.empleado.codigo or "",
                mov.fecha.isoformat() if mov.fecha else "",
                mov.fecha.day if mov.fecha else "",
                mov.tipo_movimiento_erp or "",
                mov.clave_contpaqi or "",
                _blank_if_none(mov.valor),
                _blank_if_none(mov.horas),
                _blank_if_none(mov.importe),
                mov.referencia or "",
                mov.notas or "",
            ]
        )
    return rows


def export_movimientos_contpaqi_xlsx(corte):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Movimientos_CONTPAQi"
    for row in build_movimientos_contpaqi_rows(corte):
        sheet.append(row)
    return _workbook_response(workbook, f"{corte.folio}_movimientos_contpaqi.xlsx")


def export_revision_xlsx(corte):
    workbook = Workbook()
    resumen_sheet = workbook.active
    resumen_sheet.title = "Resumen"
    resumen_sheet.append(["Campo", "Valor"])
    resumen_sheet.append(["Folio", corte.folio])
    resumen_sheet.append(["Fecha inicio", corte.fecha_inicio.isoformat() if corte.fecha_inicio else ""])
    resumen_sheet.append(["Fecha fin", corte.fecha_fin.isoformat() if corte.fecha_fin else ""])
    resumen_sheet.append(["Fecha corte", corte.fecha_corte.isoformat() if corte.fecha_corte else ""])
    resumen_sheet.append(["Estado", corte.estado])
    for key, value in sorted((corte.resumen or {}).items()):
        resumen_sheet.append([key, _blank_if_none(value)])

    empleados_sheet = workbook.create_sheet("Empleados")
    empleados_sheet.append(
        [
            "Codigo",
            "Empleado",
            "Estado",
            "Dias periodo",
            "Dias laborables",
            "Dias pre ingreso",
            "Dias asistencia",
            "Faltas",
            "Retardos",
            "Suspensiones",
            "Horas extra",
            "Ajustes pendientes",
            "Alertas bloqueantes",
            "Observaciones",
        ]
    )
    resumenes = corte.resumenes.select_related("empleado").order_by("empleado__nombre", "empleado__codigo", "id")
    for row in resumenes:
        empleados_sheet.append(
            [
                row.empleado.codigo or "",
                row.empleado.nombre or "",
                row.estado,
                row.dias_periodo,
                row.dias_laborables,
                row.dias_no_laborados_pre_ingreso,
                row.dias_asistencia,
                row.faltas,
                row.retardos,
                row.suspensiones,
                _blank_if_none(row.horas_extra_autorizadas),
                row.ajustes_pendientes,
                row.alertas_bloqueantes,
                row.observaciones or "",
            ]
        )

    movimientos_sheet = workbook.create_sheet("Movimientos_CONTPAQi")
    for row in build_movimientos_contpaqi_rows(corte):
        movimientos_sheet.append(row)

    return _workbook_response(workbook, f"{corte.folio}_revision_prenomina.xlsx")


def _workbook_response(workbook: Workbook, filename: str) -> HttpResponse:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
