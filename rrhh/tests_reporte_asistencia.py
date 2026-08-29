from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth.models import Group, User
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from rrhh.models import (
    AsistenciaEmpleado,
    Empleado,
    EmpleadoBaja,
    IncidenciaAsistencia,
    IncidenciaAsistenciaBitacora,
)
from rrhh.services_asistencia_reglas import evaluar_dia_empleado
from rrhh.views_asistencia import _empleados_reporte_asistencia


def dt_local(fecha: date, hora: time) -> datetime:
    return timezone.make_aware(datetime.combine(fecha, hora), timezone.get_current_timezone())


class ReporteAsistenciaTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="paula", password="test")
        self.user.groups.add(Group.objects.get_or_create(name="RRHH")[0])
        self.sin_permiso = User.objects.create_user(username="sinpermiso", password="test")
        self.empleado = Empleado.objects.create(
            codigo="RPT-001",
            nombre="Empleado Reporte",
            puesto="Auxiliar",
            sucursal="Matriz",
            departamento=Empleado.DEP_PRODUCCION,
            fecha_ingreso=date(2026, 6, 1),
        )
        self.fecha = date(2026, 6, 10)
        self.asistencia = AsistenciaEmpleado.objects.create(
            empleado=self.empleado,
            fecha=self.fecha,
            entrada=dt_local(self.fecha, time(8, 0)),
            salida=dt_local(self.fecha, time(16, 0)),
            minutos_comida=75,
            minutos_trabajados=480,
            fuente=AsistenciaEmpleado.FUENTE_HIKCONNECT_API,
        )
        self.incidencia = IncidenciaAsistencia.objects.create(
            empleado=self.empleado,
            fecha=self.fecha,
            tipo=IncidenciaAsistencia.TIPO_COMIDA_EXCEDIDA,
            estado=IncidenciaAsistencia.ESTADO_PENDIENTE,
            severidad=IncidenciaAsistencia.SEVERIDAD_MEDIA,
            minutos=15,
            detalle="Comida excedida por 15 minutos",
        )
        self.url = reverse("rrhh:rrhh_reporte_asistencia")
        self.editar_url = reverse("rrhh:rrhh_incidencia_editar", args=[self.incidencia.id])

    def test_query_empleados_evitar_join_cruzado(self):
        sql = str(_empleados_reporte_asistencia(self.fecha, self.fecha).query).upper()

        self.assertIn("EXISTS", sql)
        self.assertNotIn('JOIN "RRHH_ASISTENCIAEMPLEADO"', sql)
        self.assertNotIn('JOIN "RRHH_INCIDENCIAASISTENCIA"', sql)

    def test_vista_reutiliza_catalogo_empleados(self):
        self.client.force_login(self.user)

        with patch(
            "rrhh.views_asistencia._empleados_reporte_asistencia",
            wraps=_empleados_reporte_asistencia,
        ) as empleados_mock:
            response = self.client.get(
                self.url,
                {"fecha_inicio": "2026-06-10", "fecha_fin": "2026-06-10"},
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(empleados_mock.call_count, 1)
        self.assertEqual([empleado.id for empleado in response.context["empleados"]], [self.empleado.id])

    def test_vista_responde_y_resume_comida_excedida(self):
        self.client.force_login(self.user)

        response = self.client.get(
            self.url,
            {
                "fecha_inicio": "2026-06-10",
                "fecha_fin": "2026-06-10",
                "empleado": str(self.empleado.id),
            },
        )

        self.assertEqual(response.status_code, 200)
        reportes = response.context["reportes"]
        self.assertEqual(len(reportes), 1)
        self.assertEqual(reportes[0]["resumen"]["comida_excedida"], 1)
        self.assertEqual(reportes[0]["filas"][0]["incidencias"][0]["tipo"], "Comida excedida")

    def test_empleado_dado_de_baja_con_actividad_permanece_en_historial(self):
        self.empleado.activo = False
        self.empleado.save(update_fields=["activo"])
        EmpleadoBaja.objects.create(
            empleado=self.empleado,
            nombre=self.empleado.nombre,
            area=self.empleado.area,
            puesto=self.empleado.puesto,
            tipo_contrato=self.empleado.tipo_contrato,
            fecha_ingreso=self.empleado.fecha_ingreso,
            fecha_baja=date(2026, 6, 15),
            motivo=EmpleadoBaja.MOTIVO_ABANDONO,
        )
        self.client.force_login(self.user)

        response = self.client.get(
            self.url,
            {
                "fecha_inicio": "2026-06-10",
                "fecha_fin": "2026-06-10",
                "empleado": str(self.empleado.id),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual([reporte["empleado"].id for reporte in response.context["reportes"]], [self.empleado.id])
        empleados_filtro = list(response.context["empleados"])
        self.assertEqual([empleado.id for empleado in empleados_filtro], [self.empleado.id])
        self.assertEqual(empleados_filtro[0].fecha_baja_reporte, date(2026, 6, 15))
        self.assertContains(response, "Baja 2026-06-15")

    def test_empleado_dado_de_baja_sin_actividad_no_satura_el_filtro(self):
        inactivo_sin_actividad = Empleado.objects.create(
            codigo="RPT-BAJA-SIN-ACTIVIDAD",
            nombre="Empleado sin actividad en rango",
            activo=False,
            fecha_ingreso=date(2026, 1, 1),
        )
        EmpleadoBaja.objects.create(
            empleado=inactivo_sin_actividad,
            nombre=inactivo_sin_actividad.nombre,
            fecha_ingreso=inactivo_sin_actividad.fecha_ingreso,
            fecha_baja=date(2026, 5, 31),
        )
        self.client.force_login(self.user)

        response = self.client.get(
            self.url,
            {"fecha_inicio": "2026-06-10", "fecha_fin": "2026-06-10"},
        )

        self.assertNotIn(inactivo_sin_actividad.id, [empleado.id for empleado in response.context["empleados"]])

    def test_export_csv_devuelve_fila_de_incidencia(self):
        self.client.force_login(self.user)

        response = self.client.get(
            self.url,
            {
                "fecha_inicio": "2026-06-10",
                "fecha_fin": "2026-06-10",
                "empleado": str(self.empleado.id),
                "export": "csv",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response["Content-Type"].startswith("text/csv"))
        content = response.content.decode("utf-8")
        self.assertIn("RPT-001,Empleado Reporte,Matriz,2026-06-10,Comida excedida", content)
        self.assertIn("Comida excedida por 15 minutos", content)

    def test_filtro_sucursal_conserva_paridad_html_xlsx(self):
        empleado_otro = Empleado.objects.create(
            codigo="RPT-002",
            nombre="Empleado Otra Sucursal",
            sucursal="Centro",
            fecha_ingreso=date(2026, 6, 1),
        )
        IncidenciaAsistencia.objects.create(
            empleado=empleado_otro,
            fecha=self.fecha,
            tipo=IncidenciaAsistencia.TIPO_RETARDO,
            estado=IncidenciaAsistencia.ESTADO_PENDIENTE,
            severidad=IncidenciaAsistencia.SEVERIDAD_MEDIA,
            minutos=10,
            detalle="Retardo en otra sucursal",
        )
        self.client.force_login(self.user)
        params = {
            "fecha_inicio": "2026-06-10",
            "fecha_fin": "2026-06-10",
            "sucursal": "matr",
        }

        response_html = self.client.get(self.url, params)
        response_xlsx = self.client.get(self.url, {**params, "export": "xlsx"})

        self.assertEqual(response_html.status_code, 200)
        self.assertEqual(
            [reporte["empleado"].id for reporte in response_html.context["reportes"]],
            [self.empleado.id],
        )
        self.assertEqual(response_xlsx.status_code, 200)
        self.assertEqual(
            response_xlsx["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response_xlsx.content), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(rows[0][0:4], ("codigo", "nombre", "sucursal", "fecha"))
        self.assertEqual(rows[1][0:4], ("RPT-001", "Empleado Reporte", "Matriz", "2026-06-10"))
        self.assertEqual(len(rows), 2)

    def test_usuario_sin_permiso_recibe_403(self):
        self.client.force_login(self.sin_permiso)

        response = self.client.get(
            self.url,
            {
                "fecha_inicio": "2026-06-10",
                "fecha_fin": "2026-06-10",
                "empleado": str(self.empleado.id),
            },
        )

        self.assertEqual(response.status_code, 403)

    def test_edicion_con_comentario_cambia_estado_y_crea_bitacora(self):
        self.client.force_login(self.user)

        response = self.client.post(
            self.editar_url,
            {
                "estado": IncidenciaAsistencia.ESTADO_CONCILIADO,
                "minutos": str(self.incidencia.minutos),
                "detalle": self.incidencia.detalle,
                "comentario": "Validado por RRHH.",
                "fecha_inicio": "2026-06-10",
                "fecha_fin": "2026-06-10",
                "empleado": str(self.empleado.id),
                "sucursal": "Matriz",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn("fecha_inicio=2026-06-10", response["Location"])
        self.assertIn(f"empleado={self.empleado.id}", response["Location"])
        self.incidencia.refresh_from_db()
        self.assertEqual(self.incidencia.estado, IncidenciaAsistencia.ESTADO_CONCILIADO)
        self.assertTrue(self.incidencia.editado_manual)
        bitacora = IncidenciaAsistenciaBitacora.objects.get(incidencia=self.incidencia, campo="estado")
        self.assertEqual(bitacora.valor_anterior, IncidenciaAsistencia.ESTADO_PENDIENTE)
        self.assertEqual(bitacora.valor_nuevo, IncidenciaAsistencia.ESTADO_CONCILIADO)
        self.assertEqual(bitacora.comentario, "Validado por RRHH.")
        self.assertEqual(bitacora.usuario, self.user)

    def test_edicion_sin_comentario_no_cambia_ni_crea_bitacora(self):
        self.client.force_login(self.user)

        response = self.client.post(
            self.editar_url,
            {
                "estado": IncidenciaAsistencia.ESTADO_CONCILIADO,
                "minutos": str(self.incidencia.minutos),
                "detalle": self.incidencia.detalle,
                "comentario": "",
                "fecha_inicio": "2026-06-10",
                "fecha_fin": "2026-06-10",
                "empleado": str(self.empleado.id),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.incidencia.refresh_from_db()
        self.assertEqual(self.incidencia.estado, IncidenciaAsistencia.ESTADO_PENDIENTE)
        self.assertFalse(self.incidencia.editado_manual)
        self.assertFalse(IncidenciaAsistenciaBitacora.objects.filter(incidencia=self.incidencia).exists())

    def test_usuario_sin_permiso_no_puede_editar_incidencia(self):
        self.client.force_login(self.sin_permiso)

        response = self.client.post(
            self.editar_url,
            {
                "estado": IncidenciaAsistencia.ESTADO_CONCILIADO,
                "minutos": str(self.incidencia.minutos),
                "detalle": self.incidencia.detalle,
                "comentario": "Intento no autorizado.",
            },
        )

        self.assertEqual(response.status_code, 403)
        self.incidencia.refresh_from_db()
        self.assertEqual(self.incidencia.estado, IncidenciaAsistencia.ESTADO_PENDIENTE)
        self.assertFalse(IncidenciaAsistenciaBitacora.objects.filter(incidencia=self.incidencia).exists())

    def test_recalculo_no_revierte_incidencia_editada_manual(self):
        self.asistencia.salida_comida = dt_local(self.fecha, time(12, 0))
        self.asistencia.regreso_comida = dt_local(self.fecha, time(13, 15))
        self.asistencia.save(update_fields=["salida_comida", "regreso_comida"])
        self.client.force_login(self.user)
        self.client.post(
            self.editar_url,
            {
                "estado": IncidenciaAsistencia.ESTADO_RESUELTO,
                "minutos": "3",
                "detalle": "Ajuste manual validado.",
                "comentario": "Corrección manual antes de recalcular.",
                "fecha_inicio": "2026-06-10",
                "fecha_fin": "2026-06-10",
                "empleado": str(self.empleado.id),
            },
        )

        evaluar_dia_empleado(self.empleado, self.fecha)

        self.incidencia.refresh_from_db()
        self.assertEqual(self.incidencia.estado, IncidenciaAsistencia.ESTADO_RESUELTO)
        self.assertEqual(self.incidencia.minutos, 3)
        self.assertEqual(self.incidencia.detalle, "Ajuste manual validado.")
        self.assertTrue(self.incidencia.editado_manual)

    def test_falta_conciliada_suma_aparte_y_no_en_kpi_faltas(self):
        IncidenciaAsistencia.objects.create(
            empleado=self.empleado,
            fecha=self.fecha,
            tipo=IncidenciaAsistencia.TIPO_FALTA,
            estado=IncidenciaAsistencia.ESTADO_CONCILIADO,
            severidad=IncidenciaAsistencia.SEVERIDAD_INFO,
            detalle="Falta de registro conciliada con vacaciones aprobadas.",
        )
        IncidenciaAsistencia.objects.create(
            empleado=self.empleado,
            fecha=self.fecha,
            tipo=IncidenciaAsistencia.TIPO_AVISO_BAJA_FALTAS,
            estado=IncidenciaAsistencia.ESTADO_CONCILIADO,
            severidad=IncidenciaAsistencia.SEVERIDAD_INFO,
        )
        self.client.force_login(self.user)

        response = self.client.get(
            self.url,
            {
                "fecha_inicio": "2026-06-10",
                "fecha_fin": "2026-06-10",
                "empleado": str(self.empleado.id),
            },
        )

        resumen = response.context["reportes"][0]["resumen"]
        self.assertEqual(resumen["faltas"], 0)
        self.assertEqual(resumen["faltas_conciliadas"], 1)
        self.assertEqual(resumen["avisos_baja"], 0)
        # La pendiente de comida del setUp sí cuenta.
        self.assertEqual(resumen["comida_excedida"], 1)
