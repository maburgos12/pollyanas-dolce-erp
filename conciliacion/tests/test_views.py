from __future__ import annotations

from io import BytesIO
from datetime import datetime
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import load_workbook

from conciliacion.models import CfdiSucursalResolucion, ImportacionBancaria
from core.models import Sucursal
from sat_client.models import CfdiDescargado, CfdiPagoRelacionado, LogDescargaSat
from syncfy_client.models import CuentaBancaria, MovimientoBancario


class ConciliacionBancariaViewTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username="admin_conciliacion_view",
            email="admin-view@example.com",
            password="x",
        )
        self.cuenta = CuentaBancaria.objects.create(
            banco=CuentaBancaria.BANCO_BBVA,
            nombre_display="BBVA Empresas",
            id_site_syncfy="site-bbva",
            numero_cuenta="00741744000120753084",
        )
        self.client.force_login(self.user)

    def test_get_bancaria_renders_upload_screen(self):
        response = self.client.get("/conciliacion/bancaria/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Conciliacion bancaria")
        self.assertContains(response, "BBVA Empresas")
        self.assertContains(response, "Formatos aceptados: PDF, XML, CSV, XLS, XLSX o XLSM")
        self.assertNotContains(response, "accept=")

    @override_settings(SAT_DESCARGA_ENABLED=True)
    def test_get_bancaria_shows_sat_error_status_when_last_log_failed(self):
        LogDescargaSat.objects.create(nivel=LogDescargaSat.NIVEL_ERROR, mensaje="Error SAT: HTTP 500")

        response = self.client.get("/conciliacion/bancaria/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Descarga SAT con error")
        self.assertNotContains(response, "Descarga SAT activa")

    def test_get_bancaria_shows_period_bank_and_sat_summary(self):
        MovimientoBancario.objects.create(
            id_transaction="mayo-1",
            cuenta=self.cuenta,
            descripcion="DEPOSITO EN EFECTIVO MATRIZ MAYO",
            monto=Decimal("1250.00"),
            tipo=MovimientoBancario.TIPO_ABONO,
            fecha_transaccion=timezone.make_aware(datetime(2026, 5, 1, 12, 0)),
            fecha_refresh=timezone.now(),
            extra_raw={"archivo_nombre": "mayo.csv"},
        )
        MovimientoBancario.objects.create(
            id_transaction="mayo-31",
            cuenta=self.cuenta,
            descripcion="COMISION MAYO",
            monto=Decimal("15.00"),
            tipo=MovimientoBancario.TIPO_CARGO,
            fecha_transaccion=timezone.make_aware(datetime(2026, 5, 31, 12, 0)),
            fecha_refresh=timezone.now(),
            extra_raw={"archivo_nombre": "mayo.csv"},
        )
        CfdiDescargado.objects.create(
            uuid="11111111-1111-1111-1111-111111111111",
            rfc_emisor="AAA010101AAA",
            rfc_receptor="GEF211230KR2",
            subtotal=Decimal("1000.00"),
            total=Decimal("1160.00"),
            tipo_comprobante="I",
            tipo_cfdi=CfdiDescargado.TIPO_RECIBIDO,
            fecha_emision=timezone.make_aware(datetime(2026, 5, 15, 10, 0)),
        )
        matriz = Sucursal.objects.create(codigo="MATRIZ", nombre="Matriz")
        cfdi_emitido = CfdiDescargado.objects.create(
            uuid="22222222-2222-2222-2222-222222222222",
            rfc_emisor="GEF211230KR2",
            rfc_receptor="XAXX010101000",
            subtotal=Decimal("500.00"),
            total=Decimal("500.00"),
            tipo_comprobante="I",
            tipo_cfdi=CfdiDescargado.TIPO_EMITIDO,
            forma_pago="01",
            fecha_emision=timezone.make_aware(datetime(2026, 5, 16, 10, 0)),
        )
        CfdiSucursalResolucion.objects.create(
            cfdi=cfdi_emitido,
            sucursal=matriz,
            fuente=CfdiSucursalResolucion.FUENTE_XML_CONCEPTO,
            confianza=95,
            texto_detectado="VENTAS DEL DIA",
        )
        cfdi_credito = CfdiDescargado.objects.create(
            uuid="33333333-3333-3333-3333-333333333333",
            rfc_emisor="GEF211230KR2",
            rfc_receptor="CLI010101AAA",
            subtotal=Decimal("1000.00"),
            total=Decimal("1000.00"),
            tipo_comprobante="I",
            tipo_cfdi=CfdiDescargado.TIPO_EMITIDO,
            metodo_pago="PPD",
            forma_pago="99",
            fecha_emision=timezone.make_aware(datetime(2026, 4, 20, 10, 0)),
        )
        cfdi_pago = CfdiDescargado.objects.create(
            uuid="44444444-4444-4444-4444-444444444444",
            rfc_emisor="GEF211230KR2",
            rfc_receptor="CLI010101AAA",
            subtotal=Decimal("0.00"),
            total=Decimal("0.00"),
            moneda="XXX",
            tipo_comprobante="P",
            tipo_cfdi=CfdiDescargado.TIPO_EMITIDO,
            fecha_emision=timezone.make_aware(datetime(2026, 6, 3, 10, 0)),
        )
        CfdiPagoRelacionado.objects.create(
            cfdi_pago=cfdi_pago,
            uuid_relacionado=cfdi_credito.uuid,
            fecha_pago=timezone.make_aware(datetime(2026, 5, 31, 18, 45)),
            monto=Decimal("400.00"),
            moneda="MXN",
            forma_pago="03",
            num_parcialidad="1",
            importe_saldo_anterior=Decimal("1000.00"),
            importe_saldo_insoluto=Decimal("600.00"),
        )

        response = self.client.get("/conciliacion/bancaria/?periodo=2026-05")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Revision del periodo: mayo 2026")
        self.assertContains(response, "Estado del cierre")
        self.assertContains(response, "Bancos cargados")
        self.assertContains(response, "SAT del mes cargado")
        self.assertContains(response, "Matriz de reglas de conciliacion fiscal")
        self.assertContains(response, "Paquete fiscal dirigido a conciliacion")
        self.assertContains(response, "ordena ingresos por sucursal y forma de pago")
        self.assertContains(response, "Efectivo por sucursal y dia")
        self.assertContains(response, "CFDI de ingreso emitido por sucursal")
        self.assertContains(response, "factura de ingreso contra deposito bancario")
        self.assertContains(response, "Tarjetas: deposito neto")
        self.assertContains(response, "Ingreso facturado por sucursal")
        self.assertContains(response, "Complementos de pago")
        self.assertContains(response, "Resumen ejecutivo")
        self.assertContains(response, "Trabajo de conciliacion")
        self.assertContains(response, "Siguiente accion")
        self.assertContains(response, "mayo.csv")
        self.assertContains(response, "2026-05-01")
        self.assertContains(response, "2026-05-31")
        self.assertContains(response, "CFDI SAT del periodo")
        self.assertContains(response, "11111111-1111-1111-1111-111111111111")
        self.assertContains(response, "CFDI emitidos por sucursal")
        self.assertContains(response, "Matriz")
        self.assertContains(response, "$500.00")
        self.assertContains(response, "Mesa de ingresos para trabajar")
        self.assertContains(response, "Cruza lo facturado en SAT contra los abonos del banco")
        self.assertContains(response, "Depositos")
        self.assertContains(response, "Revisar diferencia")
        self.assertContains(response, "Ver depositos")
        self.assertContains(response, "Banco contra SAT por canal")
        self.assertContains(response, "Efectivo en ventanilla")
        self.assertContains(response, "Comparar factura/corte de efectivo por sucursal contra depositos")
        self.assertContains(response, "$1,250.00")
        self.assertContains(response, "$750.00")
        self.assertContains(response, "Alcance fiscal de conciliacion")
        self.assertContains(response, "Pagos cobrados del mes")
        self.assertContains(response, "$400.00")
        self.assertContains(response, "Credito clientes pendiente")
        self.assertContains(response, "$600.00")

    def test_get_bancaria_shows_all_accounts_and_filters_movements(self):
        banbajio = CuentaBancaria.objects.create(
            banco=CuentaBancaria.BANCO_BANBAJIO,
            nombre_display="BanBajio Empresas",
            id_site_syncfy="site-banbajio",
            numero_cuenta="410641890201",
        )
        amex, _ = CuentaBancaria.objects.update_or_create(
            banco=CuentaBancaria.BANCO_AMEX,
            defaults={
                "nombre_display": "American Express Business Gold",
                "id_site_syncfy": "site-amex",
                "numero_cuenta": "01005",
            },
        )
        MovimientoBancario.objects.create(
            id_transaction="banbajio-mayo",
            cuenta=banbajio,
            descripcion="DEPOSITO SUCURSAL MATRIZ",
            monto=Decimal("2500.00"),
            tipo=MovimientoBancario.TIPO_ABONO,
            fecha_transaccion=timezone.make_aware(datetime(2026, 5, 31, 12, 0)),
            fecha_refresh=timezone.now(),
        )
        MovimientoBancario.objects.create(
            id_transaction="bbva-mayo",
            cuenta=self.cuenta,
            descripcion="TRANSFERENCIA CLIENTE MAYORISTA",
            monto=Decimal("8300.00"),
            tipo=MovimientoBancario.TIPO_ABONO,
            fecha_transaccion=timezone.make_aware(datetime(2026, 5, 23, 12, 0)),
            fecha_refresh=timezone.now(),
        )
        MovimientoBancario.objects.create(
            id_transaction="amex-mayo",
            cuenta=amex,
            descripcion="AMERICAN EXPRESS CARGO",
            monto=Decimal("1900.00"),
            tipo=MovimientoBancario.TIPO_CARGO,
            fecha_transaccion=timezone.make_aware(datetime(2026, 5, 12, 12, 0)),
            fecha_refresh=timezone.now(),
        )

        response = self.client.get("/conciliacion/bancaria/?periodo=2026-05")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cuentas cargadas del periodo")
        self.assertContains(response, "BanBajio Empresas")
        self.assertContains(response, "BBVA Empresas")
        self.assertContains(response, "American Express Business Gold")
        self.assertContains(response, "Mesa de movimientos")
        self.assertContains(response, "Regla automatica")
        self.assertContains(response, "Que es este movimiento?")
        self.assertContains(response, "Ingreso ya facturado por sucursal/canal")
        self.assertContains(response, "Traspaso entre cuentas propias")
        self.assertContains(response, "Disposicion / pago linea de credito")
        self.assertContains(response, "Pago / movimiento tarjeta de credito")
        self.assertContains(response, "Cerrar como revision operativa")
        self.assertContains(response, "Sin contraparte importada / no aplica")
        self.assertContains(response, "Filtrar movimientos")
        self.assertContains(response, "Mostrando 1-3 de 3 movimientos")

        filtered = self.client.get(f"/conciliacion/bancaria/?periodo=2026-05&cuenta={amex.pk}")

        self.assertEqual(filtered.status_code, 200)
        self.assertContains(filtered, "Mostrando 1-1 de 1 movimientos")
        self.assertContains(filtered, "AMERICAN EXPRESS CARGO")
        self.assertNotContains(filtered, "TRANSFERENCIA CLIENTE MAYORISTA")

    def test_preview_and_confirm_import_movements(self):
        archivo = SimpleUploadedFile(
            "bbva.csv",
            "Fecha,Descripcion,Monto,Referencia\n2026-06-09,DEPOSITO CLIENTE,900.00,R1\n".encode("utf-8"),
            content_type="text/csv",
        )

        preview_response = self.client.post(
            "/conciliacion/bancaria/",
            {"action": "preview", "cuenta": self.cuenta.pk, "archivo": archivo},
        )
        self.assertEqual(preview_response.status_code, 200)
        self.assertContains(preview_response, "DEPOSITO CLIENTE")

        confirm_response = self.client.post("/conciliacion/bancaria/", {"action": "confirm"})

        self.assertEqual(confirm_response.status_code, 302)
        self.assertEqual(MovimientoBancario.objects.count(), 1)
        self.assertEqual(ImportacionBancaria.objects.count(), 1)

    def test_post_conciliar_movimiento_marks_credit_line_without_cfdi(self):
        movimiento = MovimientoBancario.objects.create(
            id_transaction="linea-credito-1",
            cuenta=self.cuenta,
            descripcion="DISPOSICION LINEA DE CREDITO",
            monto=Decimal("50000.00"),
            tipo=MovimientoBancario.TIPO_ABONO,
            fecha_transaccion=timezone.make_aware(datetime(2026, 5, 10, 12, 0)),
            fecha_refresh=timezone.now(),
        )

        response = self.client.post(
            "/conciliacion/bancaria/",
            {
                "action": "conciliar_movimiento",
                "periodo": "2026-05",
                "movimiento_id": str(movimiento.pk),
                "tipo_conciliacion": MovimientoBancario.CONCILIACION_LINEA_CREDITO,
                "nota_conciliacion": "Disposicion autorizada",
            },
        )

        self.assertEqual(response.status_code, 302)
        movimiento.refresh_from_db()
        self.assertTrue(movimiento.conciliado)
        self.assertEqual(movimiento.tipo_conciliacion, MovimientoBancario.CONCILIACION_LINEA_CREDITO)
        self.assertEqual(movimiento.nota_conciliacion, "Disposicion autorizada")
        self.assertEqual(movimiento.conciliado_por, self.user)

    def test_post_conciliar_movimiento_pairs_transfer_between_accounts(self):
        otra_cuenta = CuentaBancaria.objects.create(
            banco=CuentaBancaria.BANCO_BANBAJIO,
            nombre_display="BanBajio Empresas",
            id_site_syncfy="site-banbajio",
            numero_cuenta="410641890201",
        )
        cargo = MovimientoBancario.objects.create(
            id_transaction="traspaso-cargo",
            cuenta=self.cuenta,
            descripcion="TRASPASO A BANBAJIO",
            monto=Decimal("12500.00"),
            tipo=MovimientoBancario.TIPO_CARGO,
            fecha_transaccion=timezone.make_aware(datetime(2026, 5, 10, 12, 0)),
            fecha_refresh=timezone.now(),
        )
        abono = MovimientoBancario.objects.create(
            id_transaction="traspaso-abono",
            cuenta=otra_cuenta,
            descripcion="TRASPASO DE BBVA",
            monto=Decimal("12500.00"),
            tipo=MovimientoBancario.TIPO_ABONO,
            fecha_transaccion=timezone.make_aware(datetime(2026, 5, 10, 12, 5)),
            fecha_refresh=timezone.now(),
        )

        response = self.client.post(
            "/conciliacion/bancaria/",
            {
                "action": "conciliar_movimiento",
                "periodo": "2026-05",
                "movimiento_id": str(cargo.pk),
                "tipo_conciliacion": MovimientoBancario.CONCILIACION_TRASPASO,
                "contraparte_id": str(abono.pk),
                "nota_conciliacion": "Movimiento entre cuentas propias",
            },
        )

        self.assertEqual(response.status_code, 302)
        cargo.refresh_from_db()
        abono.refresh_from_db()
        self.assertTrue(cargo.conciliado)
        self.assertTrue(abono.conciliado)
        self.assertEqual(cargo.movimiento_relacionado, abono)
        self.assertEqual(abono.movimiento_relacionado, cargo)
        self.assertEqual(cargo.tipo_conciliacion, MovimientoBancario.CONCILIACION_TRASPASO)

    def test_post_conciliar_movimiento_allows_transfer_without_imported_counterpart(self):
        cargo = MovimientoBancario.objects.create(
            id_transaction="traspaso-sin-contraparte",
            cuenta=self.cuenta,
            descripcion="SPEI A CUENTA PROPIA BBVA",
            monto=Decimal("100000.00"),
            tipo=MovimientoBancario.TIPO_CARGO,
            fecha_transaccion=timezone.make_aware(datetime(2026, 5, 30, 12, 0)),
            fecha_refresh=timezone.now(),
        )

        response = self.client.post(
            "/conciliacion/bancaria/",
            {
                "action": "conciliar_movimiento",
                "periodo": "2026-05",
                "movimiento_id": str(cargo.pk),
                "tipo_conciliacion": MovimientoBancario.CONCILIACION_TRASPASO,
                "nota_conciliacion": "Traspaso a BBVA sin abono importado",
            },
        )

        self.assertEqual(response.status_code, 302)
        cargo.refresh_from_db()
        self.assertTrue(cargo.conciliado)
        self.assertEqual(cargo.tipo_conciliacion, MovimientoBancario.CONCILIACION_TRASPASO)
        self.assertIsNone(cargo.movimiento_relacionado)
        self.assertEqual(cargo.nota_conciliacion, "Traspaso a BBVA sin abono importado")

    def test_get_movimiento_detalle_shows_conciliation_document(self):
        movimiento = MovimientoBancario.objects.create(
            id_transaction="spei-propio-detalle",
            cuenta=self.cuenta,
            descripcion=(
                "SPEI Enviado: | Institucion Receptora: BBVA MEXICO | "
                "Beneficiario: GRUPO EMPRESARIAL FONSMA | Cuenta Beneficiario: 012733001207530844 "
                "Clave de Rastreo: BB2643314013215 Concepto del Pago: T"
            ),
            monto=Decimal("100000.00"),
            tipo=MovimientoBancario.TIPO_CARGO,
            fecha_transaccion=timezone.make_aware(datetime(2026, 5, 30, 12, 0)),
            fecha_refresh=timezone.now(),
            conciliado=True,
            tipo_conciliacion=MovimientoBancario.CONCILIACION_TRASPASO,
            nota_conciliacion="Traspaso a BBVA sin abono importado",
            conciliado_por=self.user,
            conciliado_en=timezone.now(),
        )

        response = self.client.get(f"/conciliacion/bancaria/movimiento/{movimiento.pk}/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Documento de conciliacion")
        self.assertContains(response, "CONC-20260530")
        self.assertContains(response, "BB2643314013215")
        self.assertContains(response, "Estado expediente")
        self.assertContains(response, "Cerrado con pendientes")
        self.assertContains(response, "Relacion contable")
        self.assertContains(response, "Traspaso entre cuentas")
        self.assertContains(response, "no se trata como gasto")
        self.assertContains(response, "Traspaso propio sin abono contraparte importado o ligado")
        self.assertContains(response, "Exportar CSV contabilidad")

    def test_get_movimiento_detalle_exports_accounting_csv(self):
        movimiento = MovimientoBancario.objects.create(
            id_transaction="spei-propio-export",
            cuenta=self.cuenta,
            descripcion="SPEI Enviado Clave de Rastreo: BB2643314013215 Concepto del Pago: T",
            monto=Decimal("100000.00"),
            tipo=MovimientoBancario.TIPO_CARGO,
            fecha_transaccion=timezone.make_aware(datetime(2026, 5, 30, 12, 0)),
            fecha_refresh=timezone.now(),
            conciliado=True,
            tipo_conciliacion=MovimientoBancario.CONCILIACION_TRASPASO,
            nota_conciliacion="Traspaso a cuenta propia",
        )

        response = self.client.get(f"/conciliacion/bancaria/movimiento/{movimiento.pk}/?export=contabilidad_csv")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        self.assertIn("contabilidad.csv", response["Content-Disposition"])
        body = response.content.decode("utf-8")
        self.assertIn("ClaveRastreo", body)
        self.assertIn("EstadoExpediente", body)
        self.assertIn("RelacionContable", body)
        self.assertIn("BB2643314013215", body)
        self.assertIn("Traspaso entre cuentas", body)

    def test_get_paquete_auditoria_shows_monthly_summary(self):
        MovimientoBancario.objects.create(
            id_transaction="paquete-mayo-1",
            cuenta=self.cuenta,
            descripcion="SPEI Enviado Clave de Rastreo: BB2643314013215 Concepto del Pago: T",
            monto=Decimal("100000.00"),
            tipo=MovimientoBancario.TIPO_CARGO,
            fecha_transaccion=timezone.make_aware(datetime(2026, 5, 30, 12, 0)),
            fecha_refresh=timezone.now(),
            conciliado=True,
            tipo_conciliacion=MovimientoBancario.CONCILIACION_TRASPASO,
            nota_conciliacion="Traspaso a cuenta propia",
        )
        CfdiDescargado.objects.create(
            uuid="55555555-5555-5555-5555-555555555555",
            rfc_emisor="GEF211230KR2",
            rfc_receptor="XAXX010101000",
            subtotal=Decimal("100.00"),
            total=Decimal("100.00"),
            tipo_comprobante="I",
            tipo_cfdi=CfdiDescargado.TIPO_EMITIDO,
            fecha_emision=timezone.make_aware(datetime(2026, 5, 30, 12, 0)),
            conciliado=True,
        )

        response = self.client.get("/conciliacion/bancaria/paquete/?periodo=2026-05")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Paquete de auditoria 2026-05")
        self.assertContains(response, "Descargar XLSX")
        self.assertContains(response, "Descargar CSV contabilidad")
        self.assertContains(response, "CSV Contabilidad Desktop")
        self.assertContains(response, "CONTPAQi Contabilidad Desktop y Nominas Desktop")
        self.assertContains(response, "Por cuenta bancaria")
        self.assertContains(response, "Excepciones y soporte pendiente")

    def test_get_paquete_auditoria_exports_xlsx(self):
        MovimientoBancario.objects.create(
            id_transaction="paquete-xlsx",
            cuenta=self.cuenta,
            descripcion="Deposito Negocios Afiliados",
            monto=Decimal("2500.00"),
            tipo=MovimientoBancario.TIPO_ABONO,
            fecha_transaccion=timezone.make_aware(datetime(2026, 5, 10, 12, 0)),
            fecha_refresh=timezone.now(),
            conciliado=True,
            tipo_conciliacion=MovimientoBancario.CONCILIACION_INGRESO_FACTURADO,
        )

        response = self.client.get("/conciliacion/bancaria/paquete/?periodo=2026-05&export=xlsx")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("paquete_conciliacion_2026-05.xlsx", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertEqual(
            workbook.sheetnames,
            [
                "Resumen",
                "Movimientos_Banco",
                "CFDI_Relacionados",
                "Poliza_Sugerida",
                "Auxiliar_Cuentas",
                "Traspasos_Propios",
                "Tarjetas_Credito",
                "Lineas_Credito",
                "Nomina",
                "Excepciones",
                "Evidencia_Pendiente",
            ],
        )
        self.assertEqual(workbook["Resumen"]["A1"].value, "Paquete mensual de conciliacion")

    def test_get_paquete_auditoria_exports_accounting_csv(self):
        MovimientoBancario.objects.create(
            id_transaction="paquete-csv",
            cuenta=self.cuenta,
            descripcion="SPEI Enviado Clave de Rastreo: BB2643314013215 Concepto del Pago: T",
            monto=Decimal("100000.00"),
            tipo=MovimientoBancario.TIPO_CARGO,
            fecha_transaccion=timezone.make_aware(datetime(2026, 5, 30, 12, 0)),
            fecha_refresh=timezone.now(),
            conciliado=True,
            tipo_conciliacion=MovimientoBancario.CONCILIACION_TRASPASO,
            extra_raw={"referencia": "2643314013215"},
        )

        response = self.client.get("/conciliacion/bancaria/paquete/?periodo=2026-05&export=contabilidad_csv")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        self.assertIn("conciliacion_2026-05_contabilidad.csv", response["Content-Disposition"])
        body = response.content.decode("utf-8")
        self.assertIn("Folio", body)
        self.assertIn("ClaveRastreo", body)
        self.assertIn("EstadoExpediente", body)
        self.assertIn("Cerrado con pendientes", body)
        self.assertIn("BB2643314013215", body)

    def test_get_paquete_auditoria_exports_contabilidad_desktop_csv(self):
        MovimientoBancario.objects.create(
            id_transaction="paquete-contabilidad-desktop",
            cuenta=self.cuenta,
            descripcion=(
                "SPEI Enviado: | Beneficiario: GRUPO EMPRESARIAL FONSMA | "
                "Cuenta Beneficiario: 012733001207530844 Clave de Rastreo: BB2643314013215 Concepto del Pago: T"
            ),
            monto=Decimal("100000.00"),
            tipo=MovimientoBancario.TIPO_CARGO,
            fecha_transaccion=timezone.make_aware(datetime(2026, 5, 30, 12, 0)),
            fecha_refresh=timezone.now(),
            conciliado=True,
            tipo_conciliacion=MovimientoBancario.CONCILIACION_TRASPASO,
            extra_raw={"referencia": "2643314013215"},
        )

        response = self.client.get("/conciliacion/bancaria/paquete/?periodo=2026-05&export=contabilidad_desktop_csv")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        self.assertIn("conciliacion_2026-05_contpaqi_contabilidad.csv", response["Content-Disposition"])
        body = response.content.decode("utf-8")
        self.assertIn("FechaPoliza", body)
        self.assertIn("TipoPolizaSugerida", body)
        self.assertIn("FolioMovimientoERP", body)
        self.assertIn("CuentaContrapartidaSugerida", body)
        self.assertIn("Cerrado con pendientes", body)
        self.assertIn("Diario", body)
        self.assertIn("BB2643314013215", body)
