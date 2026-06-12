from __future__ import annotations

import csv
from calendar import monthrange
from io import BytesIO
from datetime import date, timedelta

from django.contrib import messages
from django.conf import settings
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db.models import Count, Q, Sum
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from conciliacion.services.importador import (
    ImportacionBancariaError,
    PreviewImportacion,
    confirmar_importacion,
    generar_preview,
    periodo_default_conciliacion,
    resumen_periodo_conciliacion,
    resumen_conciliacion,
    sugerir_cfdis_para_movimientos,
)
from conciliacion.services.reglas_fiscales import regla_para_movimiento
from core.access import is_admin_or_dg
from core.audit import log_event
from sat_client.models import CfdiDescargado, LogDescargaSat
from syncfy_client.models import CuentaBancaria, MovimientoBancario


SESSION_PREVIEW_KEY = "conciliacion_bancaria_preview"
MOVIMIENTOS_PAGE_SIZE = 100


def _assert_conciliacion_access(request: HttpRequest) -> None:
    if not request.user.is_authenticated:
        raise PermissionDenied("Debes iniciar sesion.")
    if not is_admin_or_dg(request.user):
        raise PermissionDenied("No tienes permisos para conciliacion bancaria.")


@require_http_methods(["GET"])
def movimiento_conciliacion_detalle_view(request: HttpRequest, movimiento_id: int) -> HttpResponse:
    if not request.user.is_authenticated:
        return redirect(f"/login/?next=/conciliacion/bancaria/movimiento/{movimiento_id}/")
    _assert_conciliacion_access(request)
    movimiento = get_object_or_404(
        MovimientoBancario.objects.select_related(
            "cuenta",
            "cfdi_relacionado",
            "movimiento_relacionado__cuenta",
            "conciliado_por",
        ),
        pk=movimiento_id,
    )
    documento = _documento_conciliacion(movimiento)
    if request.GET.get("export") == "contabilidad_csv":
        return _export_documento_conciliacion_csv(documento)
    return render(
        request,
        "conciliacion/movimiento_detalle.html",
        {
            "documento": documento,
            "movimiento": movimiento,
        },
    )


@require_http_methods(["GET"])
def paquete_conciliacion_view(request: HttpRequest) -> HttpResponse:
    if not request.user.is_authenticated:
        return redirect("/login/?next=/conciliacion/bancaria/paquete/")
    _assert_conciliacion_access(request)
    year, month = _periodo_from_request(request)
    paquete = _paquete_conciliacion(year=year, month=month)
    export = request.GET.get("export")
    if export == "xlsx":
        return _export_paquete_xlsx(paquete)
    if export == "contabilidad_csv":
        return _export_paquete_contabilidad_csv(paquete)
    if export in {"contabilidad_desktop_csv", "contpaqi_csv"}:
        return _export_paquete_contabilidad_desktop_csv(paquete)
    return render(request, "conciliacion/paquete_auditoria.html", {"paquete": paquete})


@require_http_methods(["GET", "POST"])
def conciliacion_bancaria_view(request: HttpRequest) -> HttpResponse:
    if not request.user.is_authenticated:
        return redirect("/login/?next=/conciliacion/bancaria/")
    _assert_conciliacion_access(request)

    preview: PreviewImportacion | None = None
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "preview":
            preview = _handle_preview(request)
        elif action == "confirm":
            _handle_confirm(request)
            return redirect("conciliacion:bancaria")
        elif action == "conciliar_movimiento":
            redirect_url = _handle_conciliar_movimiento(request)
            return redirect(redirect_url)

    if preview is None:
        preview_payload = request.session.get(SESSION_PREVIEW_KEY)
        if isinstance(preview_payload, dict):
            try:
                preview = PreviewImportacion.from_session_payload(preview_payload)
            except (KeyError, TypeError, ValueError):
                request.session.pop(SESSION_PREVIEW_KEY, None)

    contexto = resumen_conciliacion()
    periodo_year, periodo_month = _periodo_from_request(request)
    periodo_resumen = resumen_periodo_conciliacion(year=periodo_year, month=periodo_month)
    movimientos_trabajo = _movimientos_trabajo_context(request, periodo_resumen)
    sat_ultimo_log = LogDescargaSat.objects.order_by("-creado_en").first()
    sat_descarga_enabled = getattr(settings, "SAT_DESCARGA_ENABLED", False)
    if sat_ultimo_log and sat_ultimo_log.nivel == LogDescargaSat.NIVEL_ERROR:
        sat_estado_label = "Descarga SAT con error"
        sat_estado_tone = "is-warn"
    elif sat_descarga_enabled:
        sat_estado_label = "Descarga SAT activa"
        sat_estado_tone = "is-ok"
    else:
        sat_estado_label = "Descarga SAT pausada"
        sat_estado_tone = "is-muted"

    contexto.update(
        {
            "cuentas": CuentaBancaria.objects.filter(activa=True).order_by("banco"),
            "sat_cfdis_total": CfdiDescargado.objects.count(),
            "sat_estado_label": sat_estado_label,
            "sat_estado_tone": sat_estado_tone,
            "sat_ultimo_log": sat_ultimo_log,
            "periodo_resumen": periodo_resumen,
            "paquete_url": f"/conciliacion/bancaria/paquete/?periodo={periodo_resumen['periodo_value']}",
            "preview": preview,
            "preview_rows": preview.movimientos[:50] if preview else [],
            "movimiento_rows": movimientos_trabajo["rows"],
            "movimientos_page": movimientos_trabajo["page"],
            "movimientos_total_filtrado": movimientos_trabajo["total"],
            "movimientos_filtros": movimientos_trabajo["filtros"],
            "movimientos_querystring": movimientos_trabajo["querystring"],
        }
    )
    return render(request, "conciliacion/bancaria.html", contexto)


def _periodo_from_request(request: HttpRequest) -> tuple[int, int]:
    periodo = str(request.GET.get("periodo") or "")
    if len(periodo) == 7 and periodo[4] == "-":
        try:
            year = int(periodo[:4])
            month = int(periodo[5:])
        except ValueError:
            return periodo_default_conciliacion()
        if 1 <= month <= 12:
            return year, month
    return periodo_default_conciliacion()


def _periodo_dates(year: int, month: int) -> tuple[date, date]:
    return date(year, month, 1), date(year, month, monthrange(year, month)[1])


def _handle_preview(request: HttpRequest) -> PreviewImportacion | None:
    cuenta_id = request.POST.get("cuenta")
    archivo = request.FILES.get("archivo")
    if not cuenta_id or not archivo:
        messages.error(request, "Selecciona una cuenta y un archivo.")
        return None
    try:
        cuenta = CuentaBancaria.objects.get(pk=cuenta_id, activa=True)
        preview = generar_preview(cuenta=cuenta, uploaded_file=archivo)
    except CuentaBancaria.DoesNotExist:
        messages.error(request, "Cuenta bancaria no valida.")
        return None
    except ImportacionBancariaError as exc:
        messages.error(request, str(exc))
        return None

    request.session[SESSION_PREVIEW_KEY] = preview.to_session_payload()
    request.session.modified = True
    messages.success(
        request,
        f"Revision lista: {len(preview.movimientos)} movimientos validos y {len(preview.errores)} filas con error.",
    )
    return preview


def _handle_confirm(request: HttpRequest) -> None:
    preview_payload = request.session.get(SESSION_PREVIEW_KEY)
    if not isinstance(preview_payload, dict):
        messages.error(request, "No hay revision pendiente para importar.")
        return
    try:
        preview = PreviewImportacion.from_session_payload(preview_payload)
        importacion = confirmar_importacion(preview=preview, user=request.user)
    except (ImportacionBancariaError, KeyError, TypeError, ValueError) as exc:
        messages.error(request, f"No se pudo confirmar la importacion: {exc}")
        return

    request.session.pop(SESSION_PREVIEW_KEY, None)
    request.session.modified = True
    log_event(
        request.user,
        "CREATE",
        "conciliacion.ImportacionBancaria",
        str(importacion.pk),
        {
            "cuenta": importacion.cuenta_id,
            "archivo_hash": importacion.archivo_hash,
            "nuevos": importacion.movimientos_nuevos,
            "duplicados": importacion.movimientos_duplicados,
        },
    )
    messages.success(
        request,
        (
            f"Importacion aplicada: {importacion.movimientos_nuevos} movimientos nuevos, "
            f"{importacion.movimientos_duplicados} duplicados."
        ),
    )


def _documento_conciliacion(movimiento: MovimientoBancario) -> dict:
    regla = regla_para_movimiento(movimiento)
    cfdi = movimiento.cfdi_relacionado
    contraparte = movimiento.movimiento_relacionado
    raw = movimiento.extra_raw or {}
    raw_payload = raw.get("raw") if isinstance(raw.get("raw"), dict) else {}
    clave_rastreo = (
        _extraer_clave_rastreo(movimiento.descripcion)
        or raw.get("referencia")
        or raw_payload.get("referencia")
        or ""
    )
    cuenta_destino = raw_payload.get("cuenta_bancaria") or _extraer_cuenta_beneficiario(movimiento.descripcion)
    evidencia_pendiente = []
    if movimiento.tipo_conciliacion == MovimientoBancario.CONCILIACION_TRASPASO and not contraparte:
        evidencia_pendiente.append("Abono contraparte no importado o no ligado.")
    if movimiento.tipo_conciliacion == MovimientoBancario.CONCILIACION_CFDI and not cfdi:
        evidencia_pendiente.append("CFDI no ligado.")
    if movimiento.tipo_conciliacion in {
        MovimientoBancario.CONCILIACION_REVISION,
        MovimientoBancario.CONCILIACION_SOPORTE,
    }:
        evidencia_pendiente.append("Requiere soporte documental externo para auditoria.")
    estado_expediente = _estado_expediente(movimiento, evidencia_pendiente)
    relacion_contable = _relacion_contable(movimiento, cfdi, contraparte, cuenta_destino)
    return {
        "id": movimiento.pk,
        "folio": f"CONC-{movimiento.fecha_transaccion:%Y%m%d}-{movimiento.pk}",
        "movimiento": movimiento,
        "regla": regla,
        "tipo_conciliacion": movimiento.get_tipo_conciliacion_display() or "Sin clasificar",
        "clave_rastreo": clave_rastreo,
        "cuenta_destino": cuenta_destino,
        "cfdi": cfdi,
        "contraparte": contraparte,
        "evidencia_pendiente": evidencia_pendiente,
        "estado_expediente": estado_expediente,
        "relacion_contable": relacion_contable,
        "contpaq": _contpaq_row(movimiento, clave_rastreo, cuenta_destino, estado_expediente, relacion_contable),
    }


def _estado_expediente(movimiento: MovimientoBancario, evidencia_pendiente: list[str]) -> dict[str, str | bool]:
    if not movimiento.conciliado:
        return {
            "label": "Pendiente de conciliar",
            "tono": "is-warn",
            "detalle": "Aun no hay conciliacion aplicada.",
            "completo": False,
        }
    if evidencia_pendiente:
        return {
            "label": "Cerrado con pendientes",
            "tono": "is-warn",
            "detalle": "Tiene clasificacion aplicada, pero falta evidencia o relacion para auditoria.",
            "completo": False,
        }
    return {
        "label": "Conciliado con expediente",
        "tono": "is-ok",
        "detalle": "La ficha muestra la relacion contable usada para cerrar el movimiento.",
        "completo": True,
    }


def _relacion_contable(
    movimiento: MovimientoBancario,
    cfdi: CfdiDescargado | None,
    contraparte: MovimientoBancario | None,
    cuenta_destino: str,
) -> dict[str, str]:
    if movimiento.tipo_conciliacion == MovimientoBancario.CONCILIACION_CFDI:
        return {
            "tipo": "Banco contra CFDI",
            "documento": cfdi.uuid if cfdi else "CFDI pendiente de ligar",
            "criterio": "Documento bancario ligado al XML fiscal que soporta ingreso, egreso o pago.",
        }
    if movimiento.tipo_conciliacion == MovimientoBancario.CONCILIACION_INGRESO_FACTURADO:
        return {
            "tipo": "Banco contra factura de ingresos",
            "documento": "CFDI de ingreso por sucursal/canal",
            "criterio": "Ingreso bancario declarado con factura de ingresos emitida por la sucursal o canal de venta.",
        }
    if movimiento.tipo_conciliacion == MovimientoBancario.CONCILIACION_TRASPASO:
        destino = cuenta_destino or (contraparte.cuenta.numero_cuenta if contraparte else "")
        return {
            "tipo": "Traspaso entre cuentas propias",
            "documento": f"Contraparte bancaria #{contraparte.pk}" if contraparte else "Contraparte no importada",
            "criterio": f"Salida y entrada entre bancos propios; no se trata como gasto. Cuenta destino: {destino or 'pendiente'}.",
        }
    if movimiento.tipo_conciliacion == MovimientoBancario.CONCILIACION_LINEA_CREDITO:
        return {
            "tipo": "Linea de credito",
            "documento": "Contrato / tabla de amortizacion / referencia bancaria",
            "criterio": "Disposicion o pago de financiamiento; se revisa contra pasivo, no contra gasto operativo.",
        }
    if movimiento.tipo_conciliacion == MovimientoBancario.CONCILIACION_TARJETA_CREDITO:
        return {
            "tipo": "Tarjeta de credito",
            "documento": "Estado de cuenta / pago de tarjeta / CFDI soporte",
            "criterio": "Movimiento de tarjeta corporativa que requiere relacion con cargos, pago o CFDI soporte.",
        }
    if movimiento.tipo_conciliacion == MovimientoBancario.CONCILIACION_COMISION:
        return {
            "tipo": "Comision bancaria",
            "documento": "Estado de cuenta / CFDI del banco cuando aplique",
            "criterio": "Cargo bancario por comision, terminal o IVA asociado.",
        }
    if movimiento.tipo_conciliacion == MovimientoBancario.CONCILIACION_FISCAL:
        return {
            "tipo": "Movimiento fiscal",
            "documento": "Declaracion / acuse / linea de captura",
            "criterio": "Pago, devolucion o ajuste fiscal revisado contra soporte SAT.",
        }
    if movimiento.tipo_conciliacion == MovimientoBancario.CONCILIACION_NOMINA:
        return {
            "tipo": "Nomina",
            "documento": "CFDI nomina / dispersion / periodo",
            "criterio": "Salida bancaria conciliada contra nomina timbrada o dispersion autorizada.",
        }
    if movimiento.tipo_conciliacion == MovimientoBancario.CONCILIACION_SOPORTE:
        return {
            "tipo": "Soporte sin CFDI",
            "documento": "Soporte externo pendiente",
            "criterio": "Clasificacion temporal; requiere documento externo para sostener auditoria.",
        }
    if movimiento.tipo_conciliacion == MovimientoBancario.CONCILIACION_REVISION:
        return {
            "tipo": "Revision operativa",
            "documento": "Revision administrativa pendiente",
            "criterio": "No debe cerrarse como conciliacion final sin documento o explicacion completa.",
        }
    return {
        "tipo": "Sin clasificar",
        "documento": "Pendiente",
        "criterio": "Seleccionar que representa el movimiento antes de cerrarlo.",
    }


def _contpaq_row(
    movimiento: MovimientoBancario,
    clave_rastreo: str,
    cuenta_destino: str,
    estado_expediente: dict[str, str | bool],
    relacion_contable: dict[str, str],
) -> dict[str, str]:
    return {
        "Fecha": movimiento.fecha_transaccion.strftime("%Y-%m-%d"),
        "Banco": movimiento.cuenta.get_banco_display(),
        "Cuenta": movimiento.cuenta.numero_cuenta or "",
        "Tipo": movimiento.tipo,
        "Importe": f"{movimiento.monto:.2f}",
        "Descripcion": movimiento.descripcion,
        "ClaveRastreo": clave_rastreo,
        "CuentaDestino": cuenta_destino,
        "ConciliacionAplicada": movimiento.get_tipo_conciliacion_display() or "",
        "EstadoExpediente": str(estado_expediente["label"]),
        "ExpedienteCompleto": "Si" if estado_expediente["completo"] else "No",
        "RelacionContable": relacion_contable["tipo"],
        "DocumentoRelacionado": relacion_contable["documento"],
        "MovimientoRelacionado": str(movimiento.movimiento_relacionado_id or ""),
        "CFDI": movimiento.cfdi_relacionado.uuid if movimiento.cfdi_relacionado_id else "",
        "Nota": movimiento.nota_conciliacion,
    }


def _export_documento_conciliacion_csv(documento: dict) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{documento["folio"]}_contabilidad.csv"'
    writer = csv.DictWriter(response, fieldnames=list(documento["contpaq"].keys()))
    writer.writeheader()
    writer.writerow(documento["contpaq"])
    return response


def _paquete_conciliacion(*, year: int, month: int) -> dict:
    inicio, fin = _periodo_dates(year, month)
    movimientos = list(
        MovimientoBancario.objects.select_related(
            "cuenta",
            "cfdi_relacionado",
            "movimiento_relacionado__cuenta",
            "conciliado_por",
        )
        .filter(fecha_transaccion__date__gte=inicio, fecha_transaccion__date__lte=fin)
        .order_by("fecha_transaccion", "cuenta__banco", "id")
    )
    documentos = [_documento_conciliacion(mov) for mov in movimientos]
    cfdis = list(
        CfdiDescargado.objects.filter(fecha_emision__date__gte=inicio, fecha_emision__date__lte=fin)
        .order_by("fecha_emision", "uuid")
    )
    movimientos_qs = MovimientoBancario.objects.filter(fecha_transaccion__date__gte=inicio, fecha_transaccion__date__lte=fin)
    cfdis_qs = CfdiDescargado.objects.filter(fecha_emision__date__gte=inicio, fecha_emision__date__lte=fin)
    excepciones = [doc for doc in documentos if doc["evidencia_pendiente"] or not doc["movimiento"].conciliado]
    return {
        "periodo_value": f"{year:04d}-{month:02d}",
        "periodo_label": inicio.strftime("%B %Y"),
        "periodo_inicio": inicio,
        "periodo_fin": fin,
        "movimientos": documentos,
        "cfdis": cfdis,
        "excepciones": excepciones,
        "resumen": {
            "movimientos_total": movimientos_qs.count(),
            "movimientos_conciliados": movimientos_qs.filter(conciliado=True).count(),
            "movimientos_pendientes": movimientos_qs.filter(conciliado=False).count(),
            "monto_total": movimientos_qs.aggregate(total=Sum("monto"))["total"] or 0,
            "cfdis_total": cfdis_qs.count(),
            "cfdis_conciliados": cfdis_qs.filter(conciliado=True).count(),
            "cfdis_pendientes": cfdis_qs.filter(conciliado=False).count(),
            "excepciones": len(excepciones),
        },
        "por_cuenta": list(
            movimientos_qs.values("cuenta__nombre_display", "cuenta__banco")
            .annotate(movimientos=Count("id"), monto=Sum("monto"), conciliados=Count("id", filter=Q(conciliado=True)))
            .order_by("cuenta__banco", "cuenta__nombre_display")
        ),
        "por_tipo": list(
            movimientos_qs.values("tipo_conciliacion")
            .annotate(movimientos=Count("id"), monto=Sum("monto"))
            .order_by("tipo_conciliacion")
        ),
    }


def _export_paquete_contabilidad_csv(paquete: dict) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="conciliacion_{paquete["periodo_value"]}_contabilidad.csv"'
    fieldnames = list(_paquete_movimiento_row(paquete["movimientos"][0]).keys()) if paquete["movimientos"] else ["Periodo"]
    writer = csv.DictWriter(response, fieldnames=fieldnames)
    writer.writeheader()
    for documento in paquete["movimientos"]:
        writer.writerow(_paquete_movimiento_row(documento))
    return response


def _export_paquete_contabilidad_desktop_csv(paquete: dict) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="conciliacion_{paquete["periodo_value"]}_contpaqi_contabilidad.csv"'
    rows = [_poliza_sugerida_row(documento) for documento in paquete["movimientos"]]
    fieldnames = list(rows[0].keys()) if rows else ["Periodo"]
    writer = csv.DictWriter(response, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    return response


def _export_paquete_xlsx(paquete: dict) -> HttpResponse:
    workbook = Workbook()
    resumen = workbook.active
    resumen.title = "Resumen"
    _write_key_values(
        resumen,
        [
            ("Periodo", paquete["periodo_value"]),
            ("Fecha inicial", paquete["periodo_inicio"].isoformat()),
            ("Fecha final", paquete["periodo_fin"].isoformat()),
            ("Movimientos", paquete["resumen"]["movimientos_total"]),
            ("Conciliados", paquete["resumen"]["movimientos_conciliados"]),
            ("Pendientes", paquete["resumen"]["movimientos_pendientes"]),
            ("CFDI SAT", paquete["resumen"]["cfdis_total"]),
            ("CFDI pendientes", paquete["resumen"]["cfdis_pendientes"]),
            ("Excepciones / soporte pendiente", paquete["resumen"]["excepciones"]),
        ],
    )
    _write_table(resumen, 13, ["Cuenta", "Banco", "Movimientos", "Conciliados", "Monto"], [
        [
            row["cuenta__nombre_display"],
            row["cuenta__banco"],
            row["movimientos"],
            row["conciliados"],
            row["monto"],
        ]
        for row in paquete["por_cuenta"]
    ])

    movimientos_sheet = workbook.create_sheet("Movimientos_Banco")
    movimiento_rows = [_paquete_movimiento_row(documento) for documento in paquete["movimientos"]]
    _write_dict_table(movimientos_sheet, movimiento_rows)

    cfdi_sheet = workbook.create_sheet("CFDI_Relacionados")
    _write_dict_table(cfdi_sheet, [_paquete_cfdi_row(cfdi) for cfdi in paquete["cfdis"]])

    poliza_sheet = workbook.create_sheet("Poliza_Sugerida")
    _write_dict_table(poliza_sheet, [_poliza_sugerida_row(documento) for documento in paquete["movimientos"]])

    auxiliar_sheet = workbook.create_sheet("Auxiliar_Cuentas")
    _write_dict_table(auxiliar_sheet, [_auxiliar_cuenta_row(documento) for documento in paquete["movimientos"]])

    traspasos_sheet = workbook.create_sheet("Traspasos_Propios")
    _write_dict_table(
        traspasos_sheet,
        [
            _traspaso_propio_row(documento)
            for documento in paquete["movimientos"]
            if documento["movimiento"].tipo_conciliacion == MovimientoBancario.CONCILIACION_TRASPASO
        ],
    )

    tarjetas_sheet = workbook.create_sheet("Tarjetas_Credito")
    _write_dict_table(
        tarjetas_sheet,
        [
            _instrumento_financiero_row(documento)
            for documento in paquete["movimientos"]
            if documento["movimiento"].tipo_conciliacion == MovimientoBancario.CONCILIACION_TARJETA_CREDITO
        ],
    )

    lineas_sheet = workbook.create_sheet("Lineas_Credito")
    _write_dict_table(
        lineas_sheet,
        [
            _instrumento_financiero_row(documento)
            for documento in paquete["movimientos"]
            if documento["movimiento"].tipo_conciliacion == MovimientoBancario.CONCILIACION_LINEA_CREDITO
        ],
    )

    nomina_sheet = workbook.create_sheet("Nomina")
    _write_dict_table(
        nomina_sheet,
        [
            _nomina_row(documento)
            for documento in paquete["movimientos"]
            if documento["movimiento"].tipo_conciliacion == MovimientoBancario.CONCILIACION_NOMINA
        ],
    )

    excepciones_sheet = workbook.create_sheet("Excepciones")
    excepcion_rows = [_paquete_excepcion_row(documento) for documento in paquete["excepciones"]]
    _write_dict_table(excepciones_sheet, excepcion_rows)

    evidencia_sheet = workbook.create_sheet("Evidencia_Pendiente")
    _write_dict_table(evidencia_sheet, [_evidencia_pendiente_row(documento) for documento in paquete["excepciones"]])

    for sheet in workbook.worksheets:
        _autosize_sheet(sheet)
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="paquete_conciliacion_{paquete["periodo_value"]}.xlsx"'
    return response


def _paquete_movimiento_row(documento: dict) -> dict[str, str]:
    movimiento = documento["movimiento"]
    return {
        "Folio": documento["folio"],
        "Fecha": movimiento.fecha_transaccion.strftime("%Y-%m-%d"),
        "Banco": movimiento.cuenta.get_banco_display(),
        "Cuenta": movimiento.cuenta.numero_cuenta or "",
        "TipoBanco": movimiento.tipo,
        "Importe": f"{movimiento.monto:.2f}",
        "DescripcionBanco": movimiento.descripcion,
        "ClaveRastreo": documento["clave_rastreo"],
        "CuentaDestino": documento["cuenta_destino"],
        "ReglaAutomatica": documento["regla"].mesa_label,
        "ConciliacionAplicada": documento["tipo_conciliacion"],
        "EstadoExpediente": documento["estado_expediente"]["label"],
        "ExpedienteCompleto": "Si" if documento["estado_expediente"]["completo"] else "No",
        "RelacionContable": documento["relacion_contable"]["tipo"],
        "DocumentoRelacionado": documento["relacion_contable"]["documento"],
        "MovimientoRelacionado": str(movimiento.movimiento_relacionado_id or ""),
        "CFDI": movimiento.cfdi_relacionado.uuid if movimiento.cfdi_relacionado_id else "",
        "PendientesEvidencia": " | ".join(documento["evidencia_pendiente"]),
        "Nota": movimiento.nota_conciliacion,
    }


def _paquete_excepcion_row(documento: dict) -> dict[str, str]:
    row = _paquete_movimiento_row(documento)
    row["MotivoExcepcion"] = " | ".join(documento["evidencia_pendiente"]) or "Movimiento no conciliado"
    return row


def _paquete_cfdi_row(cfdi: CfdiDescargado) -> dict[str, str]:
    return {
        "UUID": cfdi.uuid,
        "Fecha": cfdi.fecha_emision.strftime("%Y-%m-%d"),
        "TipoCFDI": cfdi.get_tipo_cfdi_display(),
        "TipoComprobante": cfdi.tipo_comprobante,
        "RFCEmisor": cfdi.rfc_emisor,
        "NombreEmisor": cfdi.nombre_emisor or "",
        "RFCReceptor": cfdi.rfc_receptor,
        "NombreReceptor": cfdi.nombre_receptor or "",
        "Total": f"{cfdi.total:.2f}",
        "Conciliado": "Si" if cfdi.conciliado else "No",
    }


def _poliza_sugerida_row(documento: dict) -> dict[str, str]:
    movimiento = documento["movimiento"]
    return {
        "FechaPoliza": movimiento.fecha_transaccion.strftime("%Y-%m-%d"),
        "TipoPolizaSugerida": _tipo_poliza_sugerida(movimiento),
        "FolioMovimientoERP": documento["folio"],
        "ConceptoPoliza": f"{documento['tipo_conciliacion']} | {documento['clave_rastreo'] or movimiento.descripcion[:40]}",
        "CargoBanco": f"{movimiento.monto:.2f}" if movimiento.tipo == MovimientoBancario.TIPO_ABONO else "0.00",
        "AbonoBanco": f"{movimiento.monto:.2f}" if movimiento.tipo == MovimientoBancario.TIPO_CARGO else "0.00",
        "CuentaContableBanco": "",
        "CuentaContrapartidaSugerida": _cuenta_contrapartida_sugerida(movimiento),
        "UUIDRelacionado": movimiento.cfdi_relacionado.uuid if movimiento.cfdi_relacionado_id else "",
        "RFCRelacionado": _rfc_relacionado(movimiento.cfdi_relacionado) if movimiento.cfdi_relacionado_id else "",
        "MetodoPago": _metodo_pago_contpaqi(movimiento),
        "Referencia": documento["clave_rastreo"],
        "EstadoExpediente": documento["estado_expediente"]["label"],
        "EvidenciaPendiente": " | ".join(documento["evidencia_pendiente"]),
        "Nota": movimiento.nota_conciliacion,
    }


def _auxiliar_cuenta_row(documento: dict) -> dict[str, str]:
    movimiento = documento["movimiento"]
    return {
        "Fecha": movimiento.fecha_transaccion.strftime("%Y-%m-%d"),
        "CuentaContableBanco": "",
        "CuentaBanco": movimiento.cuenta.numero_cuenta or "",
        "Banco": movimiento.cuenta.get_banco_display(),
        "CuentaContrapartidaSugerida": _cuenta_contrapartida_sugerida(movimiento),
        "TipoMovimiento": movimiento.tipo,
        "Cargo": f"{movimiento.monto:.2f}" if movimiento.tipo == MovimientoBancario.TIPO_CARGO else "0.00",
        "Abono": f"{movimiento.monto:.2f}" if movimiento.tipo == MovimientoBancario.TIPO_ABONO else "0.00",
        "RelacionContable": documento["relacion_contable"]["tipo"],
        "DocumentoRelacionado": documento["relacion_contable"]["documento"],
        "UUIDRelacionado": movimiento.cfdi_relacionado.uuid if movimiento.cfdi_relacionado_id else "",
        "Referencia": documento["clave_rastreo"],
        "EstadoExpediente": documento["estado_expediente"]["label"],
    }


def _traspaso_propio_row(documento: dict) -> dict[str, str]:
    movimiento = documento["movimiento"]
    contraparte = documento["contraparte"]
    return {
        "Folio": documento["folio"],
        "Fecha": movimiento.fecha_transaccion.strftime("%Y-%m-%d"),
        "CuentaSalida": movimiento.cuenta.numero_cuenta or "",
        "CuentaEntrada": documento["cuenta_destino"] or (contraparte.cuenta.numero_cuenta if contraparte else ""),
        "BancoSalida": movimiento.cuenta.get_banco_display(),
        "BancoEntrada": contraparte.cuenta.get_banco_display() if contraparte else "",
        "Importe": f"{movimiento.monto:.2f}",
        "ClaveRastreo": documento["clave_rastreo"],
        "MovimientoContraparte": str(contraparte.pk if contraparte else ""),
        "Pendiente": "Abono contraparte no importado" if not contraparte else "",
        "Descripcion": movimiento.descripcion,
    }


def _instrumento_financiero_row(documento: dict) -> dict[str, str]:
    movimiento = documento["movimiento"]
    return {
        "Folio": documento["folio"],
        "Fecha": movimiento.fecha_transaccion.strftime("%Y-%m-%d"),
        "Tipo": documento["tipo_conciliacion"],
        "Banco": movimiento.cuenta.get_banco_display(),
        "CuentaPago": movimiento.cuenta.numero_cuenta or "",
        "Importe": f"{movimiento.monto:.2f}",
        "Referencia": documento["clave_rastreo"],
        "CuentaContrapartidaSugerida": _cuenta_contrapartida_sugerida(movimiento),
        "DocumentoSoporte": documento["relacion_contable"]["documento"],
        "EstadoExpediente": documento["estado_expediente"]["label"],
        "EvidenciaPendiente": " | ".join(documento["evidencia_pendiente"]),
        "Descripcion": movimiento.descripcion,
    }


def _nomina_row(documento: dict) -> dict[str, str]:
    movimiento = documento["movimiento"]
    return {
        "Folio": documento["folio"],
        "FechaPago": movimiento.fecha_transaccion.strftime("%Y-%m-%d"),
        "Banco": movimiento.cuenta.get_banco_display(),
        "Cuenta": movimiento.cuenta.numero_cuenta or "",
        "ImporteDispersado": f"{movimiento.monto:.2f}",
        "CFDINomina": movimiento.cfdi_relacionado.uuid if movimiento.cfdi_relacionado_id else "",
        "PeriodoNomina": "",
        "Referencia": documento["clave_rastreo"],
        "EstadoExpediente": documento["estado_expediente"]["label"],
        "EvidenciaPendiente": " | ".join(documento["evidencia_pendiente"]),
        "Descripcion": movimiento.descripcion,
    }


def _evidencia_pendiente_row(documento: dict) -> dict[str, str]:
    movimiento = documento["movimiento"]
    return {
        "Folio": documento["folio"],
        "Fecha": movimiento.fecha_transaccion.strftime("%Y-%m-%d"),
        "Banco": movimiento.cuenta.get_banco_display(),
        "Cuenta": movimiento.cuenta.numero_cuenta or "",
        "Importe": f"{movimiento.monto:.2f}",
        "ConciliacionAplicada": documento["tipo_conciliacion"],
        "EstadoExpediente": documento["estado_expediente"]["label"],
        "Pendiente": " | ".join(documento["evidencia_pendiente"]) or "Movimiento no conciliado",
        "DocumentoRelacionado": documento["relacion_contable"]["documento"],
        "Referencia": documento["clave_rastreo"],
    }


def _metodo_pago_contpaqi(movimiento: MovimientoBancario) -> str:
    if "SPEI" in (movimiento.descripcion or "").upper() or movimiento.tipo_conciliacion == MovimientoBancario.CONCILIACION_TRASPASO:
        return "Transferencia"
    if movimiento.tipo_conciliacion == MovimientoBancario.CONCILIACION_TARJETA_CREDITO:
        return "Tarjeta credito"
    return ""


def _tipo_poliza_sugerida(movimiento: MovimientoBancario) -> str:
    if movimiento.tipo_conciliacion == MovimientoBancario.CONCILIACION_TRASPASO:
        return "Diario"
    if movimiento.tipo == MovimientoBancario.TIPO_ABONO:
        return "Ingreso"
    return "Egreso"


def _rfc_relacionado(cfdi: CfdiDescargado) -> str:
    return cfdi.rfc_receptor if cfdi.tipo_cfdi == CfdiDescargado.TIPO_EMITIDO else cfdi.rfc_emisor


def _cuenta_contrapartida_sugerida(movimiento: MovimientoBancario) -> str:
    mapping = {
        MovimientoBancario.CONCILIACION_TRASPASO: "Banco contraparte",
        MovimientoBancario.CONCILIACION_INGRESO_FACTURADO: "Ingresos por ventas",
        MovimientoBancario.CONCILIACION_COMISION: "Comisiones bancarias",
        MovimientoBancario.CONCILIACION_FISCAL: "Impuestos",
        MovimientoBancario.CONCILIACION_NOMINA: "Nomina",
        MovimientoBancario.CONCILIACION_TARJETA_CREDITO: "Tarjeta credito",
        MovimientoBancario.CONCILIACION_CFDI: "Proveedor/cliente CFDI",
    }
    return mapping.get(movimiento.tipo_conciliacion, "Revision contable")


def _write_key_values(sheet, rows: list[tuple[str, object]]) -> None:
    sheet["A1"] = "Paquete mensual de conciliacion"
    sheet["A1"].font = Font(bold=True, size=14)
    for idx, (key, value) in enumerate(rows, start=3):
        sheet.cell(row=idx, column=1, value=key).font = Font(bold=True)
        sheet.cell(row=idx, column=2, value=value)


def _write_table(sheet, start_row: int, headers: list[str], rows: list[list[object]]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2E8F0")
    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)


def _write_dict_table(sheet, rows: list[dict[str, str]]) -> None:
    if not rows:
        sheet["A1"] = "Sin registros"
        return
    headers = list(rows[0].keys())
    _write_table(sheet, 1, headers, [[row.get(header, "") for header in headers] for row in rows])


def _autosize_sheet(sheet) -> None:
    for column in sheet.columns:
        max_length = 0
        letter = get_column_letter(column[0].column)
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 48)


def _extraer_clave_rastreo(descripcion: str) -> str:
    marker = "Clave de Rastreo:"
    if marker not in descripcion:
        return ""
    return descripcion.split(marker, 1)[1].split("|", 1)[0].split("Concepto", 1)[0].strip()


def _extraer_cuenta_beneficiario(descripcion: str) -> str:
    marker = "Cuenta Beneficiario:"
    if marker not in descripcion:
        return ""
    return descripcion.split(marker, 1)[1].split(" ", 1)[0].strip()


def _handle_conciliar_movimiento(request: HttpRequest) -> str:
    periodo = str(request.POST.get("periodo") or "").strip()
    redirect_url = f"/conciliacion/bancaria/?periodo={periodo}#mesa-movimientos" if periodo else "/conciliacion/bancaria/#mesa-movimientos"
    movimiento_id = str(request.POST.get("movimiento_id") or "").strip()
    tipo_conciliacion = str(request.POST.get("tipo_conciliacion") or "").strip()
    nota = str(request.POST.get("nota_conciliacion") or "").strip()

    if tipo_conciliacion not in dict(MovimientoBancario.CONCILIACION_CHOICES):
        messages.error(request, "Selecciona una accion de conciliacion valida.")
        return redirect_url
    try:
        movimiento = MovimientoBancario.objects.select_related("cuenta").get(pk=movimiento_id)
    except (MovimientoBancario.DoesNotExist, ValueError):
        messages.error(request, "Movimiento bancario no encontrado.")
        return redirect_url

    if tipo_conciliacion == MovimientoBancario.CONCILIACION_CFDI:
        cfdi_uuid = str(request.POST.get("cfdi_uuid") or "").strip()
        try:
            cfdi = CfdiDescargado.objects.get(uuid=cfdi_uuid)
        except CfdiDescargado.DoesNotExist:
            messages.error(request, "Selecciona un CFDI valido para relacionar.")
            return redirect_url
        _marcar_movimiento_conciliado(
            movimiento,
            tipo_conciliacion=tipo_conciliacion,
            user=request.user,
            nota=nota,
            cfdi=cfdi,
        )
        cfdi.conciliado = True
        cfdi.save(update_fields=["conciliado"])
        messages.success(request, "Movimiento conciliado contra CFDI.")
        return redirect_url

    if tipo_conciliacion == MovimientoBancario.CONCILIACION_TRASPASO:
        contraparte_id = str(request.POST.get("contraparte_id") or "").strip()
        if not contraparte_id:
            _marcar_movimiento_conciliado(
                movimiento,
                tipo_conciliacion=tipo_conciliacion,
                user=request.user,
                nota=nota or "Traspaso entre cuentas propias sin contraparte importada o localizada.",
            )
            messages.success(request, "Movimiento marcado como traspaso entre cuentas sin contraparte ligada.")
            return redirect_url
        try:
            contraparte = MovimientoBancario.objects.select_related("cuenta").get(pk=contraparte_id)
        except (MovimientoBancario.DoesNotExist, ValueError):
            messages.error(request, "Selecciona una contraparte valida o deja el campo vacio si no esta importada.")
            return redirect_url
        if not _es_contraparte_valida(movimiento, contraparte):
            messages.error(request, "La contraparte debe ser otro movimiento con monto igual y tipo opuesto.")
            return redirect_url
        _marcar_movimiento_conciliado(
            movimiento,
            tipo_conciliacion=tipo_conciliacion,
            user=request.user,
            nota=nota,
            movimiento_relacionado=contraparte,
        )
        _marcar_movimiento_conciliado(
            contraparte,
            tipo_conciliacion=tipo_conciliacion,
            user=request.user,
            nota=nota or f"Contraparte de movimiento {movimiento.pk}",
            movimiento_relacionado=movimiento,
        )
        messages.success(request, "Traspaso entre cuentas conciliado con su contraparte.")
        return redirect_url

    _marcar_movimiento_conciliado(
        movimiento,
        tipo_conciliacion=tipo_conciliacion,
        user=request.user,
        nota=nota,
    )
    messages.success(request, "Movimiento clasificado y marcado como conciliado.")
    return redirect_url


def _marcar_movimiento_conciliado(
    movimiento: MovimientoBancario,
    *,
    tipo_conciliacion: str,
    user,
    nota: str = "",
    cfdi: CfdiDescargado | None = None,
    movimiento_relacionado: MovimientoBancario | None = None,
) -> None:
    movimiento.conciliado = True
    movimiento.tipo_conciliacion = tipo_conciliacion
    movimiento.nota_conciliacion = nota
    movimiento.cfdi_relacionado = cfdi
    movimiento.movimiento_relacionado = movimiento_relacionado
    movimiento.conciliado_por = user if getattr(user, "is_authenticated", False) else None
    movimiento.conciliado_en = timezone.now()
    movimiento.save(
        update_fields=[
            "conciliado",
            "tipo_conciliacion",
            "nota_conciliacion",
            "cfdi_relacionado",
            "movimiento_relacionado",
            "conciliado_por",
            "conciliado_en",
        ]
    )


def _es_contraparte_valida(movimiento: MovimientoBancario, contraparte: MovimientoBancario) -> bool:
    return (
        movimiento.pk != contraparte.pk
        and movimiento.tipo != contraparte.tipo
        and movimiento.monto == contraparte.monto
    )


def _movimientos_trabajo_context(request: HttpRequest, periodo_resumen: dict) -> dict:
    qs = MovimientoBancario.objects.select_related(
        "cuenta",
        "cfdi_relacionado",
        "movimiento_relacionado__cuenta",
        "conciliado_por",
    ).filter(
        fecha_transaccion__date__gte=periodo_resumen["periodo_inicio"],
        fecha_transaccion__date__lte=periodo_resumen["periodo_fin"],
    )
    cuenta = str(request.GET.get("cuenta") or "").strip()
    tipo = str(request.GET.get("tipo") or "").strip()
    busqueda = str(request.GET.get("q") or "").strip()

    if cuenta and cuenta.isdigit():
        qs = qs.filter(cuenta_id=cuenta)
    else:
        cuenta = ""
    if tipo in {MovimientoBancario.TIPO_ABONO, MovimientoBancario.TIPO_CARGO}:
        qs = qs.filter(tipo=tipo)
    else:
        tipo = ""
    if busqueda:
        qs = qs.filter(
            Q(descripcion__icontains=busqueda)
            | Q(cuenta__nombre_display__icontains=busqueda)
            | Q(cuenta__numero_cuenta__icontains=busqueda)
        )

    qs = qs.order_by("-fecha_transaccion", "cuenta__banco", "id")
    paginator = Paginator(qs, MOVIMIENTOS_PAGE_SIZE)
    page = paginator.get_page(request.GET.get("page") or 1)
    movimientos = list(page.object_list)
    candidatos = sugerir_cfdis_para_movimientos([mov for mov in movimientos if not mov.conciliado])
    query = request.GET.copy()
    query.pop("page", None)
    return {
        "rows": _movimiento_rows(movimientos, candidatos),
        "page": page,
        "total": paginator.count,
        "filtros": {"cuenta": cuenta, "tipo": tipo, "q": busqueda},
        "querystring": query.urlencode(),
    }


def _movimiento_rows(movimientos, candidatos: dict[int, list]) -> list[dict]:
    rows = []
    contrapartes = _contrapartes_por_movimiento(movimientos)
    for movimiento in movimientos:
        rows.append(
            {
                "movimiento": movimiento,
                "candidatos": candidatos.get(movimiento.pk, []),
                "contrapartes": contrapartes.get(movimiento.pk, []),
                "regla": regla_para_movimiento(movimiento),
                "documento": _documento_conciliacion(movimiento),
            }
        )
    return rows


def _contrapartes_por_movimiento(movimientos) -> dict[int, list[MovimientoBancario]]:
    result = {}
    for movimiento in movimientos:
        opposite = MovimientoBancario.TIPO_ABONO if movimiento.tipo == MovimientoBancario.TIPO_CARGO else MovimientoBancario.TIPO_CARGO
        start = movimiento.fecha_transaccion - timedelta(days=5)
        end = movimiento.fecha_transaccion + timedelta(days=5)
        result[movimiento.pk] = list(
            MovimientoBancario.objects.select_related("cuenta")
            .filter(
                conciliado=False,
                tipo=opposite,
                monto=movimiento.monto,
                fecha_transaccion__gte=start,
                fecha_transaccion__lte=end,
            )
            .exclude(pk=movimiento.pk)
            .order_by("fecha_transaccion", "cuenta__banco")[:5]
        )
    return result
