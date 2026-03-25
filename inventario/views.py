import csv
import os
import subprocess
import sys
from datetime import timedelta
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
from urllib.parse import quote_plus, urlencode

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.core.files.uploadedfile import UploadedFile
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from core.access import (
    ROLE_ADMIN,
    ROLE_DG,
    can_manage_inventario,
    can_view_inventario,
    can_view_maestros,
    has_any_role,
)
from core.audit import log_event
from compras.models import SolicitudCompra
from maestros.models import CostoInsumo, Insumo, InsumoAlias, PointPendingMatch, Proveedor
from maestros.utils.canonical_catalog import (
    canonical_insumo_by_id,
    canonicalized_active_insumos,
    canonicalized_insumo_selector,
    enterprise_readiness_profile,
    usage_maps_for_insumo_ids,
)
from recetas.models import LineaReceta, Receta, RecetaCodigoPointAlias, VentaHistorica, normalizar_codigo_point
from recetas.utils.matching import clasificar_match, match_insumo
from recetas.utils.normalizacion import normalizar_nombre
from inventario.utils.almacen_import import (
    ENTRADAS_FILE,
    INVENTARIO_FILE,
    MERMA_FILE,
    SALIDAS_FILE,
    import_folder,
)
from inventario.utils.google_drive_sync import get_drive_sync_mode, sync_almacen_from_drive
from inventario.utils.sync_logging import log_sync_run
from inventario.utils.reorder import FORMULA_EXCEL_LEGACY, FORMULA_LEADTIME_PLUS_SAFETY, calcular_punto_reorden

from .models import AjusteInventario, AlmacenSyncRun, ExistenciaInsumo, InventarioConfig, MovimientoInventario


SOURCE_TO_FILENAME = {
    "inventario": INVENTARIO_FILE,
    "entradas": ENTRADAS_FILE,
    "salidas": SALIDAS_FILE,
    "merma": MERMA_FILE,
}

FILENAME_TO_SOURCE = {v: k for k, v in SOURCE_TO_FILENAME.items()}


def _canonicalized_insumo_stock_options(limit: int = 200) -> list[dict]:
    canonical_rows = canonicalized_active_insumos(limit=limit)
    grouped_member_ids = [member_id for row in canonical_rows for member_id in row["member_ids"]]
    usage_maps = usage_maps_for_insumo_ids(grouped_member_ids)
    existencias_by_insumo = {
        row["insumo_id"]: row["stock_actual"]
        for row in ExistenciaInsumo.objects.filter(insumo_id__in=grouped_member_ids).values("insumo_id", "stock_actual")
    }
    options = []
    for row in canonical_rows:
        insumo = row["canonical"]
        stock_total = sum((existencias_by_insumo.get(member_id, Decimal("0")) for member_id in row["member_ids"]), Decimal("0"))
        enterprise_profile = getattr(insumo, "enterprise_profile", enterprise_readiness_profile(insumo))
        movement_count = sum(int(usage_maps["movement_counts"].get(member_id, 0)) for member_id in row["member_ids"])
        adjustment_count = sum(int(usage_maps["adjustment_counts"].get(member_id, 0)) for member_id in row["member_ids"])
        has_existence = any(member_id in usage_maps["existence_ids"] for member_id in row["member_ids"])
        inventory_refs = movement_count + adjustment_count + (1 if has_existence else 0)
        is_operational_blocker = enterprise_profile["readiness_label"] == "Incompleto" and inventory_refs > 0
        options.append(
            {
                "id": insumo.id,
                "nombre": insumo.nombre,
                "stock": stock_total,
                "canonical_variant_count": row["variant_count"],
                "enterprise_status": enterprise_profile["readiness_label"],
                "enterprise_missing": enterprise_profile["missing"],
                "is_operational_blocker": is_operational_blocker,
                "operational_blocker_label": "Bloquea inventario" if is_operational_blocker else "",
            }
        )
    return options


def _canonicalized_existencias_rows(limit: int = 200) -> list[SimpleNamespace]:
    canonical_rows = canonicalized_active_insumos(limit=limit)
    member_ids = [member_id for row in canonical_rows for member_id in row["member_ids"]]
    usage_maps = usage_maps_for_insumo_ids(member_ids)
    existencias = {
        existencia.insumo_id: existencia
        for existencia in ExistenciaInsumo.objects.filter(insumo_id__in=member_ids).select_related("insumo", "insumo__unidad_base")
    }
    rows: list[SimpleNamespace] = []
    formula_mode = getattr(settings, "INVENTARIO_REORDER_FORMULA", FORMULA_EXCEL_LEGACY)

    for row in canonical_rows:
        insumo = row["canonical"]
        member_existencias = [existencias[member_id] for member_id in row["member_ids"] if member_id in existencias]
        canonical_existencia = existencias.get(insumo.id)
        base_existencia = canonical_existencia or (member_existencias[0] if member_existencias else None)

        stock_actual = sum((Decimal(str(item.stock_actual or 0)) for item in member_existencias), Decimal("0"))
        stock_minimo = Decimal(str(getattr(base_existencia, "stock_minimo", 0) or 0))
        stock_maximo = Decimal(str(getattr(base_existencia, "stock_maximo", 0) or 0))
        punto_reorden = Decimal(str(getattr(base_existencia, "punto_reorden", 0) or 0))
        inventario_promedio = Decimal(str(getattr(base_existencia, "inventario_promedio", 0) or 0))
        dias_llegada_pedido = int(getattr(base_existencia, "dias_llegada_pedido", 0) or 0)
        consumo_diario_promedio = Decimal(str(getattr(base_existencia, "consumo_diario_promedio", 0) or 0))
        recomendado = calcular_punto_reorden(
            stock_minimo=stock_minimo,
            dias_llegada_pedido=dias_llegada_pedido,
            consumo_diario_promedio=consumo_diario_promedio,
            formula=formula_mode,
        )
        enterprise_profile = getattr(insumo, "enterprise_profile", enterprise_readiness_profile(insumo))
        movement_count = sum(int(usage_maps["movement_counts"].get(member_id, 0)) for member_id in row["member_ids"])
        adjustment_count = sum(int(usage_maps["adjustment_counts"].get(member_id, 0)) for member_id in row["member_ids"])
        has_existence = any(member_id in usage_maps["existence_ids"] for member_id in row["member_ids"])
        inventory_refs = movement_count + adjustment_count + (1 if has_existence else 0)
        is_operational_blocker = enterprise_profile["readiness_label"] == "Incompleto" and inventory_refs > 0
        usage_label = (
            "Producción interna"
            if insumo.tipo_item == Insumo.TIPO_INTERNO
            else "Empaque final"
            if insumo.tipo_item == Insumo.TIPO_EMPAQUE
            else "Compra directa"
        )
        primary_missing = enterprise_profile["missing"][0] if enterprise_profile["missing"] else ""
        missing_field = (
            "unidad"
            if primary_missing == "unidad base"
            else "proveedor"
            if primary_missing == "proveedor principal"
            else "categoria"
            if primary_missing == "categoría"
            else "codigo_point"
            if primary_missing == "código comercial"
            else None
        )
        edit_query = {"insumo_id": insumo.id, "usage_scope": "inventory"}
        if missing_field:
            edit_query["missing_field"] = missing_field

        rows.append(
            SimpleNamespace(
                insumo=insumo,
                stock_actual=stock_actual,
                stock_minimo=stock_minimo,
                stock_maximo=stock_maximo,
                punto_reorden=punto_reorden,
                inventario_promedio=inventario_promedio,
                dias_llegada_pedido=dias_llegada_pedido,
                consumo_diario_promedio=consumo_diario_promedio,
                punto_reorden_recomendado=recomendado,
                punto_reorden_diferencia=punto_reorden - recomendado,
                canonical_variant_count=row["variant_count"],
                enterprise_profile=enterprise_profile,
                enterprise_status=enterprise_profile["readiness_label"],
                enterprise_missing=enterprise_profile["missing"],
                enterprise_usage_label=usage_label,
                enterprise_primary_missing=primary_missing,
                enterprise_missing_field=missing_field,
                inventory_refs=inventory_refs,
                canonical_pending=row["variant_count"] > 1 and inventory_refs > 0,
                canonical_pending_label="Consolidación pendiente" if row["variant_count"] > 1 and inventory_refs > 0 else "",
                canonical_list_url=reverse("maestros:insumo_list")
                + f"?{urlencode({'usage_scope': 'inventory', 'canonical_status': 'variantes', 'q': insumo.nombre})}",
                is_operational_blocker=is_operational_blocker,
                operational_blocker_label="Bloquea inventario" if is_operational_blocker else "",
                enterprise_edit_url=reverse("maestros:insumo_update", args=[insumo.id]),
                enterprise_list_url=reverse("maestros:insumo_list") + f"?{urlencode(edit_query)}",
            )
        )
    return rows


def _inventory_master_blocker_context(
    rows: list[SimpleNamespace],
    *,
    usage_scope: str = "inventory",
    focus_summary_template: str,
    row_action_detail: str,
    card_action_detail: str,
    current_view_url: str = "",
    current_query: dict[str, object] | None = None,
    selected_focus_key: str = "auto",
) -> dict[str, object]:
    blocker_rows = [row for row in rows if getattr(row, "is_operational_blocker", False)]
    blocker_groups: dict[str, dict[str, object]] = {}
    for row in blocker_rows:
        class_key = row.insumo.tipo_item or Insumo.TIPO_MATERIA_PRIMA
        class_label = (
            "Insumo interno"
            if class_key == Insumo.TIPO_INTERNO
            else "Empaque"
            if class_key == Insumo.TIPO_EMPAQUE
            else "Materia prima"
        )
        group = blocker_groups.setdefault(
            class_key,
            {
                "class_key": class_key,
                "class_label": class_label,
                "count": 0,
                "missing_totals": defaultdict(int),
            },
        )
        group["count"] += 1
        for missing_label in row.enterprise_missing:
            group["missing_totals"][missing_label] += 1

    blocker_cards: list[dict[str, object]] = []
    blocker_detail_rows: list[dict[str, object]] = []
    base_query = {str(k): v for k, v in (current_query or {}).items() if v not in (None, "", [], ())}
    for group in sorted(blocker_groups.values(), key=lambda item: (-int(item["count"]), str(item["class_label"]))):
        dominant_label = ""
        dominant_count = 0
        for missing_label, count in dict(group["missing_totals"]).items():
            if count > dominant_count:
                dominant_label = str(missing_label)
                dominant_count = int(count)
        query = {
            "tipo_item": group["class_key"],
            "enterprise_status": "incompletos",
            "usage_scope": usage_scope,
        }
        filter_key = (
            "unidad"
            if dominant_label == "unidad base"
            else "proveedor"
            if dominant_label == "proveedor principal"
            else "categoria"
            if dominant_label == "categoría"
            else "codigo_point"
            if dominant_label == "código comercial"
            else None
        )
        if filter_key:
            query["missing_field"] = filter_key
        focus_key = f"{group['class_key']}:{filter_key or 'all'}"
        focus_query = {**base_query, "master_focus_key": focus_key}
        blocker_cards.append(
            {
                "class_label": group["class_label"],
                "count": group["count"],
                "dominant_label": dominant_label or "maestro incompleto",
                "dominant_count": dominant_count,
                "action_label": "Corregir maestro",
                "action_detail": card_action_detail,
                "action_url": reverse("maestros:insumo_list") + f"?{urlencode(query)}",
                "focus_key": focus_key,
                "focus_url": f"{current_view_url}?{urlencode(focus_query)}" if current_view_url else "",
                "is_active": selected_focus_key == focus_key,
            }
        )

    for row in blocker_rows[:8]:
        primary_missing = row.enterprise_missing[0] if row.enterprise_missing else "maestro"
        filter_key = (
            "unidad"
            if primary_missing == "unidad base"
            else "proveedor"
            if primary_missing == "proveedor principal"
            else "categoria"
            if primary_missing == "categoría"
            else "codigo_point"
            if primary_missing == "código comercial"
            else None
        )
        query = {
            "enterprise_status": "incompletos",
            "usage_scope": usage_scope,
            "insumo_id": row.insumo.id,
        }
        if filter_key:
            query["missing_field"] = filter_key
        blocker_detail_rows.append(
            {
                "class_label": (
                    "Insumo interno"
                    if row.insumo.tipo_item == Insumo.TIPO_INTERNO
                    else "Empaque"
                    if row.insumo.tipo_item == Insumo.TIPO_EMPAQUE
                    else "Materia prima"
                ),
                "name": row.insumo.nombre,
                "missing_field": filter_key or "maestro",
                "missing": ", ".join(row.enterprise_missing),
                "action_label": "Corregir artículo",
                "action_detail": row_action_detail,
                "action_url": reverse("maestros:insumo_list") + f"?{urlencode(query)}",
                "edit_url": reverse("maestros:insumo_update", args=[row.insumo.id]),
                "tone": "warning",
                "focus_key": f"{row.insumo.tipo_item or Insumo.TIPO_MATERIA_PRIMA}:{filter_key or 'all'}",
            }
        )

    active_focus_key = selected_focus_key if selected_focus_key and selected_focus_key != "auto" else ""
    filtered_focus_rows = [row for row in blocker_detail_rows if not active_focus_key or row.get("focus_key") == active_focus_key]
    master_focus_rows = list((filtered_focus_rows or blocker_detail_rows)[:8])
    if master_focus_rows:
        first_master_focus = master_focus_rows[0]
        master_focus = {
            **first_master_focus,
            "label": f"{first_master_focus['class_label']} · {first_master_focus['missing_field']}",
            "summary": focus_summary_template.format(
                name=first_master_focus["name"],
                missing_field=first_master_focus["missing_field"],
            ),
            "tone": "warning",
        }
    else:
        master_focus = {
            "class_label": "Maestro ERP",
            "missing_field": "sin bloqueos",
            "label": "Maestro ERP al día",
            "summary": "No hay artículos bloqueando inventario por faltantes del maestro.",
            "action_label": "Abrir maestro",
            "action_detail": "Puedes revisar el catálogo para depuración preventiva.",
            "action_url": reverse("maestros:insumo_list"),
            "tone": "success",
            "count": 0,
        }

    return {
        "master_blocker_class_cards": blocker_cards,
        "master_blocker_detail_rows": blocker_detail_rows,
        "master_blocker_total": len(blocker_rows),
        "master_focus": master_focus,
        "master_focus_rows": master_focus_rows,
        "selected_master_focus_key": active_focus_key,
    }


def _map_alias_import_header(raw: str) -> str:
    h = normalizar_nombre(raw)
    if h in {
        "alias",
        "nombre",
        "nombre origen",
        "nombre_origen",
        "origen",
        "insumo origen",
        "insumo_origen",
        "nombre almacen",
        "nombre_almacen",
        "point nombre",
        "point_nombre",
    }:
        return "alias"
    if h in {
        "insumo",
        "insumo oficial",
        "insumo_oficial",
        "insumo destino",
        "insumo_destino",
        "canonico",
        "oficial",
        "insumo erp",
        "insumo_erp",
    }:
        return "insumo"
    return h


def _read_alias_import_rows(uploaded: UploadedFile) -> list[dict]:
    ext = Path(uploaded.name or "").suffix.lower()
    rows: list[dict] = []

    if ext in {".xlsx", ".xlsm"}:
        uploaded.seek(0)
        wb = load_workbook(uploaded, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        first_row = next(rows_iter, None)
        if not first_row:
            return []
        headers = [_map_alias_import_header(str(h or "")) for h in first_row]
        for raw in rows_iter:
            row = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                row[header] = raw[idx] if idx < len(raw) else None
            rows.append(row)
        return rows

    if ext == ".csv":
        uploaded.seek(0)
        content = uploaded.read().decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(StringIO(content))
        for raw in reader:
            row = {}
            for k, v in raw.items():
                if not k:
                    continue
                row[_map_alias_import_header(k)] = v
            rows.append(row)
        return rows

    raise ValueError("Formato no soportado. Usa .xlsx, .xlsm o .csv.")


def _to_decimal(value: str, default: str = "0") -> Decimal:
    try:
        return Decimal(value or default)
    except (InvalidOperation, TypeError):
        return Decimal(default)


def _inventario_reorder_max_diff_pct() -> Decimal:
    default_pct = _to_decimal(str(getattr(settings, "INVENTARIO_REORDER_MAX_DIFF_PCT", 10)), "10")
    return InventarioConfig.get_solo(default_pct=default_pct).reorder_max_diff_pct


def _apply_movimiento(movimiento: MovimientoInventario) -> None:
    insumo_canonical = canonical_insumo_by_id(movimiento.insumo_id) or movimiento.insumo
    existencia, _ = ExistenciaInsumo.objects.get_or_create(insumo=insumo_canonical)
    if movimiento.tipo == MovimientoInventario.TIPO_ENTRADA:
        existencia.stock_actual += movimiento.cantidad
    else:
        existencia.stock_actual -= movimiento.cantidad
    existencia.actualizado_en = timezone.now()
    existencia.save()


def _can_approve_ajustes(user) -> bool:
    return has_any_role(user, ROLE_ADMIN, ROLE_DG)


def _apply_ajuste(ajuste: AjusteInventario, acted_by, comentario: str = "") -> None:
    insumo_canonical = canonical_insumo_by_id(ajuste.insumo_id) or ajuste.insumo
    existencia, _ = ExistenciaInsumo.objects.get_or_create(insumo=insumo_canonical)
    prev_stock = existencia.stock_actual
    delta = ajuste.cantidad_fisica - ajuste.cantidad_sistema
    if ajuste.insumo_id == insumo_canonical.id:
        existencia.stock_actual = ajuste.cantidad_fisica
    else:
        existencia.stock_actual = prev_stock + delta
    existencia.actualizado_en = timezone.now()
    existencia.save()
    log_event(
        acted_by,
        "APPLY",
        "inventario.ExistenciaInsumo",
        existencia.id,
        {
            "source": ajuste.folio,
            "insumo_id": insumo_canonical.id,
            "from_stock": str(prev_stock),
            "to_stock": str(existencia.stock_actual),
            "delta": str(delta),
        },
    )

    if delta != 0:
        movimiento_ajuste = MovimientoInventario.objects.create(
            tipo=MovimientoInventario.TIPO_ENTRADA if delta > 0 else MovimientoInventario.TIPO_SALIDA,
            insumo=insumo_canonical,
            cantidad=abs(delta),
            referencia=ajuste.folio,
        )
        log_event(
            acted_by,
            "CREATE",
            "inventario.MovimientoInventario",
            movimiento_ajuste.id,
            {
                "tipo": movimiento_ajuste.tipo,
                "insumo_id": insumo_canonical.id,
                "cantidad": str(movimiento_ajuste.cantidad),
                "referencia": movimiento_ajuste.referencia,
            },
        )

    ajuste.estatus = AjusteInventario.STATUS_APLICADO
    ajuste.aprobado_por = acted_by if acted_by and acted_by.is_authenticated else None
    now = timezone.now()
    ajuste.aprobado_en = now
    ajuste.aplicado_en = now
    ajuste.comentario_revision = (comentario or ajuste.comentario_revision or "")[:255]
    ajuste.save(update_fields=["estatus", "aprobado_por", "aprobado_en", "aplicado_en", "comentario_revision"])
    log_event(
        acted_by,
        "APPLY",
        "inventario.AjusteInventario",
        ajuste.id,
        {
            "folio": ajuste.folio,
            "estatus": ajuste.estatus,
            "comentario_revision": ajuste.comentario_revision,
        },
    )


def _classify_upload(filename: str) -> tuple[str | None, str | None]:
    if not filename:
        return None, None
    original = filename.strip()
    normalized = normalizar_nombre(Path(original).name)

    for expected_name, source in FILENAME_TO_SOURCE.items():
        if normalized == normalizar_nombre(expected_name):
            return source, expected_name

    if "inventario" in normalized and "almacen" in normalized:
        return "inventario", INVENTARIO_FILE
    if "entradas" in normalized and "almacen" in normalized:
        return "entradas", ENTRADAS_FILE
    if "salidas" in normalized and "almacen" in normalized:
        return "salidas", SALIDAS_FILE
    if "merma" in normalized and "almacen" in normalized:
        return "merma", MERMA_FILE
    return None, None


def _save_uploaded_file(target_dir: str, target_name: str, uploaded: UploadedFile) -> None:
    filepath = Path(target_dir) / target_name
    with filepath.open("wb") as out:
        for chunk in uploaded.chunks():
            out.write(chunk)


def _upsert_alias(alias_name: str, insumo: Insumo) -> tuple[bool, str, str]:
    alias_norm = normalizar_nombre(alias_name)
    if not alias_norm:
        return False, "", "El nombre origen no es válido."

    obj, created = InsumoAlias.objects.get_or_create(
        nombre_normalizado=alias_norm,
        defaults={
            "nombre": alias_name[:250],
            "insumo": insumo,
        },
    )
    if not created and (obj.insumo_id != insumo.id or obj.nombre != alias_name[:250]):
        obj.insumo = insumo
        obj.nombre = alias_name[:250]
        obj.save(update_fields=["insumo", "nombre"])
    return True, alias_norm, "creado" if created else "actualizado"


def _remove_pending_name_from_session(request: HttpRequest, alias_norm: str) -> None:
    _remove_pending_names_from_session(request, {alias_norm})


def _remove_pending_name_from_recent_runs(alias_norm: str, max_runs: int = 20) -> None:
    _remove_pending_names_from_recent_runs({alias_norm}, max_runs=max_runs)


def _remove_pending_names_from_session(request: HttpRequest, alias_norms: set[str]) -> None:
    if not alias_norms:
        return
    pending = request.session.get("inventario_pending_preview")
    if not pending:
        return
    request.session["inventario_pending_preview"] = [
        row
        for row in pending
        if normalizar_nombre(str((row or {}).get("nombre_origen") or "")) not in alias_norms
    ]


def _remove_pending_names_from_recent_runs(alias_norms: set[str], max_runs: int = 20) -> None:
    if not alias_norms:
        return
    runs = AlmacenSyncRun.objects.only("id", "pending_preview").order_by("-started_at")[:max_runs]
    for run in runs:
        pending = list(getattr(run, "pending_preview", []) or [])
        filtered = [
            row
            for row in pending
            if normalizar_nombre(str((row or {}).get("nombre_origen") or "")) not in alias_norms
        ]
        if len(filtered) != len(pending):
            run.pending_preview = filtered
            run.save(update_fields=["pending_preview"])


def _build_pending_grouped(pending_preview: list[dict]) -> list[dict]:
    grouped = defaultdict(lambda: {"count": 0, "sources": set(), "name": "", "suggestion": "", "score_max": 0.0})
    for row in pending_preview:
        name = str((row or {}).get("nombre_origen") or "").strip()
        norm = str((row or {}).get("nombre_normalizado") or normalizar_nombre(name))
        if not norm:
            continue
        item = grouped[norm]
        item["count"] += 1
        item["name"] = item["name"] or name
        source = str((row or {}).get("source") or "").strip()
        if source:
            item["sources"].add(source)
        suggestion = str((row or {}).get("sugerencia") or "").strip()
        if suggestion and not item["suggestion"]:
            item["suggestion"] = suggestion
        try:
            score = float((row or {}).get("score") or 0.0)
        except (TypeError, ValueError):
            score = 0.0
        item["score_max"] = max(item["score_max"], score)

    return sorted(
        [
            {
                "nombre_origen": v["name"],
                "nombre_normalizado": k,
                "count": v["count"],
                "sources": ", ".join(sorted(v["sources"])) if v["sources"] else "-",
                "sugerencia": v["suggestion"],
                "score_max": v["score_max"],
            }
            for k, v in grouped.items()
        ],
        key=lambda x: (-x["count"], x["nombre_origen"]),
    )


def _build_cross_unified_rows(
    pending_grouped: list[dict],
    *,
    point_tipos: list[str] | None = None,
) -> tuple[list[dict], int, int]:
    unified = defaultdict(
        lambda: {
            "nombre_muestra": "",
            "point_count": 0,
            "almacen_count": 0,
            "receta_count": 0,
            "suggestion": "",
            "score_max": 0.0,
        }
    )

    for row in pending_grouped:
        norm = row["nombre_normalizado"]
        if not norm:
            continue
        item = unified[norm]
        item["nombre_muestra"] = item["nombre_muestra"] or row["nombre_origen"]
        item["almacen_count"] += int(row["count"] or 0)
        if row.get("sugerencia") and not item["suggestion"]:
            item["suggestion"] = row["sugerencia"]
        item["score_max"] = max(item["score_max"], float(row.get("score_max") or 0.0))

    point_pending_qs = PointPendingMatch.qs_operativos()
    if point_tipos is None:
        point_pending_qs = point_pending_qs.filter(tipo=PointPendingMatch.TIPO_INSUMO)
    else:
        point_pending_qs = point_pending_qs.filter(tipo__in=point_tipos)
    point_pending_matches = list(point_pending_qs.only("point_nombre", "fuzzy_sugerencia", "fuzzy_score"))
    for pending in point_pending_matches:
        norm = normalizar_nombre(pending.point_nombre or "")
        if not norm:
            continue
        item = unified[norm]
        item["nombre_muestra"] = item["nombre_muestra"] or (pending.point_nombre or "")
        item["point_count"] += 1
        if (pending.fuzzy_sugerencia or "").strip() and not item["suggestion"]:
            item["suggestion"] = pending.fuzzy_sugerencia.strip()
        item["score_max"] = max(item["score_max"], float(pending.fuzzy_score or 0.0))

    receta_pending_lines_qs = LineaReceta.objects.filter(~Q(tipo_linea=LineaReceta.TIPO_SUBSECCION)).filter(
        Q(insumo__isnull=True) | Q(match_status=LineaReceta.STATUS_NEEDS_REVIEW) | Q(match_status=LineaReceta.STATUS_REJECTED)
    )
    receta_pending_lines = 0
    for linea in receta_pending_lines_qs.only("insumo_texto"):
        norm = normalizar_nombre(linea.insumo_texto or "")
        if not norm:
            continue
        item = unified[norm]
        item["nombre_muestra"] = item["nombre_muestra"] or (linea.insumo_texto or "")
        item["receta_count"] += 1
        receta_pending_lines += 1

    unified_rows = []
    overlaps = 0
    for norm, item in unified.items():
        sources_active = sum(
            1
            for value in (item["point_count"], item["almacen_count"], item["receta_count"])
            if value > 0
        )
        if sources_active >= 2:
            overlaps += 1
        unified_rows.append(
            {
                "nombre_normalizado": norm,
                "nombre_muestra": item["nombre_muestra"] or norm,
                "point_count": item["point_count"],
                "almacen_count": item["almacen_count"],
                "receta_count": item["receta_count"],
                "sources_active": sources_active,
                "total_count": item["point_count"] + item["almacen_count"] + item["receta_count"],
                "suggestion": item["suggestion"],
                "score_max": item["score_max"],
            }
        )
    unified_rows.sort(key=lambda x: (-x["sources_active"], -x["total_count"], x["nombre_muestra"]))
    return unified_rows, len(point_pending_matches), receta_pending_lines


def _read_cross_filters(params) -> tuple[str, str, bool, int, float]:
    cross_q = (params.get("cross_q") or "").strip()
    cross_q_norm = normalizar_nombre(cross_q)
    cross_only_suggested = (params.get("cross_only_suggested") or "").strip().lower() in {"1", "true", "on", "yes"}
    cross_min_sources = int(_to_decimal(params.get("cross_min_sources"), "1"))
    cross_min_sources = max(1, min(3, cross_min_sources))
    cross_score_min = float(_to_decimal(params.get("cross_score_min"), "0"))
    cross_score_min = max(0.0, min(100.0, cross_score_min))
    return cross_q, cross_q_norm, cross_only_suggested, cross_min_sources, cross_score_min


def _read_cross_point_tipo(params) -> tuple[str, list[str] | None]:
    point_tipo = (params.get("cross_point_tipo") or PointPendingMatch.TIPO_INSUMO).strip().upper()
    valid_point_tipos = {
        PointPendingMatch.TIPO_INSUMO,
        PointPendingMatch.TIPO_PROVEEDOR,
        PointPendingMatch.TIPO_PRODUCTO,
        "TODOS",
        "ALL",
    }
    if point_tipo not in valid_point_tipos:
        point_tipo = PointPendingMatch.TIPO_INSUMO
    point_tipos_filter = None if point_tipo in {"TODOS", "ALL"} else [point_tipo]
    return point_tipo, point_tipos_filter


def _read_cross_source(params) -> str:
    cross_source = (params.get("cross_source") or "TODOS").strip().upper()
    valid_sources = {"TODOS", "ALL", "ALMACEN", "POINT", "RECETAS"}
    if cross_source not in valid_sources:
        cross_source = "TODOS"
    return cross_source


def _read_cross_table_controls(params) -> tuple[int, int, str, str]:
    cross_limit = int(_to_decimal(params.get("cross_limit"), "120"))
    cross_limit = max(1, min(500, cross_limit))
    cross_offset = int(_to_decimal(params.get("cross_offset"), "0"))
    cross_offset = max(0, min(50000, cross_offset))

    cross_sort_by = (params.get("cross_sort_by") or "sources_active").strip().lower()
    allowed_sort = {
        "sources_active",
        "total_count",
        "score_max",
        "point_count",
        "almacen_count",
        "receta_count",
        "nombre_muestra",
        "nombre_normalizado",
    }
    if cross_sort_by not in allowed_sort:
        cross_sort_by = "sources_active"

    cross_sort_dir = (params.get("cross_sort_dir") or "desc").strip().lower()
    if cross_sort_dir not in {"asc", "desc"}:
        cross_sort_dir = "desc"

    return cross_limit, cross_offset, cross_sort_by, cross_sort_dir


def _sort_cross_rows(cross_rows: list[dict], *, sort_by: str, sort_dir: str) -> list[dict]:
    allowed_sort = {
        "sources_active": lambda row: int(row.get("sources_active") or 0),
        "total_count": lambda row: int(row.get("total_count") or 0),
        "score_max": lambda row: float(row.get("score_max") or 0.0),
        "point_count": lambda row: int(row.get("point_count") or 0),
        "almacen_count": lambda row: int(row.get("almacen_count") or 0),
        "receta_count": lambda row: int(row.get("receta_count") or 0),
        "nombre_muestra": lambda row: str(row.get("nombre_muestra") or "").lower(),
        "nombre_normalizado": lambda row: str(row.get("nombre_normalizado") or "").lower(),
    }
    sort_key = allowed_sort.get(sort_by, allowed_sort["sources_active"])
    reverse = sort_dir == "desc"
    return sorted(
        cross_rows,
        key=lambda row: (sort_key(row), str(row.get("nombre_muestra") or "").lower()),
        reverse=reverse,
    )


def _apply_cross_filters(
    unified_rows: list[dict],
    cross_q_norm: str,
    cross_only_suggested: bool,
    cross_min_sources: int,
    cross_score_min: float,
) -> list[dict]:
    filtered_rows = []
    for row in unified_rows:
        sources_active = int(row.get("sources_active") or 0)
        score_max = float(row.get("score_max") or 0.0)
        suggestion = str(row.get("suggestion") or "").strip()
        nombre_muestra_norm = normalizar_nombre(str(row.get("nombre_muestra") or ""))
        nombre_norm = str(row.get("nombre_normalizado") or "")
        suggestion_norm = normalizar_nombre(suggestion)

        if sources_active < cross_min_sources:
            continue
        if score_max < cross_score_min:
            continue
        if cross_only_suggested and not suggestion:
            continue
        if cross_q_norm and (cross_q_norm not in nombre_muestra_norm) and (cross_q_norm not in nombre_norm) and (
            cross_q_norm not in suggestion_norm
        ):
            continue
        filtered_rows.append(row)
    return filtered_rows


def _load_visible_pending_preview(request: HttpRequest, max_rows: int = 120, max_runs: int = 20) -> list[dict]:
    session_pending = list(request.session.get("inventario_pending_preview", []))[:max_rows]
    if session_pending:
        return session_pending

    hidden_run_id = request.session.get("inventario_hidden_pending_run_id")
    latest_runs = list(AlmacenSyncRun.objects.only("id", "pending_preview").order_by("-started_at")[:max_runs])
    latest_pending_run = None
    for run in latest_runs:
        if hidden_run_id and run.id == hidden_run_id:
            continue
        if isinstance(run.pending_preview, list) and run.pending_preview:
            latest_pending_run = run
            break
    return list((latest_pending_run.pending_preview if latest_pending_run else [])[:max_rows])


def _export_cross_pending_csv(cross_unified_rows: list[dict]) -> HttpResponse:
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="inventario_catalogo_pendientes_{now_str}.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "nombre_muestra",
            "nombre_normalizado",
            "point_count",
            "almacen_count",
            "receta_count",
            "fuentes_activas",
            "total_count",
            "sugerencia",
            "score_max",
        ]
    )
    for row in cross_unified_rows:
        writer.writerow(
            [
                row.get("nombre_muestra", ""),
                row.get("nombre_normalizado", ""),
                row.get("point_count", 0),
                row.get("almacen_count", 0),
                row.get("receta_count", 0),
                row.get("sources_active", 0),
                row.get("total_count", 0),
                row.get("suggestion", ""),
                row.get("score_max", 0.0),
            ]
        )
    return response


def _export_cross_pending_xlsx(cross_unified_rows: list[dict]) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "catalogo_pendientes"
    ws.append(
        [
            "nombre_muestra",
            "nombre_normalizado",
            "point_count",
            "almacen_count",
            "receta_count",
            "fuentes_activas",
            "total_count",
            "sugerencia",
            "score_max",
        ]
    )
    for row in cross_unified_rows:
        ws.append(
            [
                row.get("nombre_muestra", ""),
                row.get("nombre_normalizado", ""),
                row.get("point_count", 0),
                row.get("almacen_count", 0),
                row.get("receta_count", 0),
                row.get("sources_active", 0),
                row.get("total_count", 0),
                row.get("suggestion", ""),
                row.get("score_max", 0.0),
            ]
        )
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 34
    ws.column_dimensions["I"].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="inventario_catalogo_pendientes_{now_str}.xlsx"'
    return response


def _export_alias_template(export_format: str) -> HttpResponse:
    headers = ["alias", "insumo"]
    sample_rows = [
        ["Harina pastelera 25kg", "Harina Pastelera"],
        ["Mantequilla barra", "Mantequilla"],
        ["Fresa fresca premium", "Fresa Fresca"],
    ]

    if export_format == "alias_template_csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="plantilla_aliases_inventario.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(sample_rows)
        return response

    wb = Workbook()
    ws = wb.active
    ws.title = "aliases_import"
    ws.append(headers)
    for row in sample_rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 36

    bytes_buffer = BytesIO()
    wb.save(bytes_buffer)
    bytes_buffer.seek(0)
    response = HttpResponse(
        bytes_buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_aliases_inventario.xlsx"'
    return response


def _export_aliases_catalog_csv(aliases_qs) -> HttpResponse:
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="inventario_aliases_catalogo_{now_str}.csv"'
    writer = csv.writer(response)
    writer.writerow(["alias", "normalizado", "insumo_oficial"])
    for alias in aliases_qs:
        writer.writerow([alias.nombre, alias.nombre_normalizado, alias.insumo.nombre if alias.insumo_id else ""])
    return response


def _export_aliases_catalog_xlsx(aliases_qs) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "aliases_catalogo"
    ws.append(["alias", "normalizado", "insumo_oficial"])
    for alias in aliases_qs:
        ws.append([alias.nombre, alias.nombre_normalizado, alias.insumo.nombre if alias.insumo_id else ""])
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 36
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="inventario_aliases_catalogo_{now_str}.xlsx"'
    return response


def _export_alias_import_preview_csv(preview_rows: list[dict]) -> HttpResponse:
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="inventario_aliases_import_preview_{now_str}.csv"'
    writer = csv.writer(response)
    writer.writerow(["row", "alias", "insumo_archivo", "sugerencia", "score", "method", "motivo"])
    for row in preview_rows:
        writer.writerow(
            [
                row.get("row", ""),
                row.get("alias", ""),
                row.get("insumo_archivo", ""),
                row.get("sugerencia", ""),
                row.get("score", ""),
                row.get("method", ""),
                row.get("motivo", ""),
            ]
        )
    return response


def _export_alias_import_preview_xlsx(preview_rows: list[dict]) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "aliases_import_preview"
    ws.append(["row", "alias", "insumo_archivo", "sugerencia", "score", "method", "motivo"])
    for row in preview_rows:
        ws.append(
            [
                row.get("row", ""),
                row.get("alias", ""),
                row.get("insumo_archivo", ""),
                row.get("sugerencia", ""),
                row.get("score", ""),
                row.get("method", ""),
                row.get("motivo", ""),
            ]
        )
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 40
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    now_str = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="inventario_aliases_import_preview_{now_str}.xlsx"'
    return response


def _resolve_cross_source_with_alias(alias_name: str, insumo: Insumo) -> tuple[int, int]:
    alias_norm = normalizar_nombre(alias_name)
    if not alias_norm:
        return 0, 0

    point_resolved = 0
    for pending in PointPendingMatch.qs_operativos().filter(tipo=PointPendingMatch.TIPO_INSUMO).only("id", "point_nombre", "point_codigo"):
        if normalizar_nombre(pending.point_nombre or "") != alias_norm:
            continue
        point_code = (pending.point_codigo or "").strip()
        changed = []
        if point_code and (insumo.codigo_point or "").strip() != point_code:
            insumo.codigo_point = point_code[:80]
            changed.append("codigo_point")
        if (pending.point_nombre or "").strip() and insumo.nombre_point != pending.point_nombre:
            insumo.nombre_point = pending.point_nombre[:250]
            changed.append("nombre_point")
        if changed:
            insumo.save(update_fields=changed)
        pending.delete()
        point_resolved += 1

    canonical_row = next((row for row in canonicalized_active_insumos(limit=5000) if insumo.id in row["member_ids"]), None)
    canonical_member_ids = canonical_row["member_ids"] if canonical_row else [insumo.id]
    latest_cost = (
        CostoInsumo.objects.filter(insumo_id__in=canonical_member_ids)
        .order_by("-fecha", "-id")
        .values_list("costo_unitario", flat=True)
        .first()
    )
    recetas_resolved = 0
    lineas_qs = LineaReceta.objects.filter(~Q(tipo_linea=LineaReceta.TIPO_SUBSECCION)).filter(
        Q(insumo__isnull=True) | Q(match_status=LineaReceta.STATUS_NEEDS_REVIEW) | Q(match_status=LineaReceta.STATUS_REJECTED)
    )
    for linea in lineas_qs.only("id", "insumo_texto", "unidad_texto", "unidad_id", "cantidad"):
        if normalizar_nombre(linea.insumo_texto or "") != alias_norm:
            continue
        linea.insumo = insumo
        if not linea.unidad_id and insumo.unidad_base_id:
            linea.unidad = insumo.unidad_base
        if (not linea.unidad_texto) and insumo.unidad_base_id and insumo.unidad_base:
            linea.unidad_texto = insumo.unidad_base.codigo
        if latest_cost is not None:
            linea.costo_unitario_snapshot = latest_cost
        linea.match_status = LineaReceta.STATUS_AUTO
        linea.match_method = "ALIAS"
        linea.match_score = 100.0
        linea.save(
            update_fields=[
                "insumo",
                "unidad",
                "unidad_texto",
                "costo_unitario_snapshot",
                "match_status",
                "match_method",
                "match_score",
            ]
        )
        recetas_resolved += 1

    return point_resolved, recetas_resolved


def _latest_cost_by_insumo_cached(
    insumo_id: int,
    cache: dict[int, Decimal | None],
    member_ids_cache: dict[int, list[int]] | None = None,
) -> Decimal | None:
    if insumo_id in cache:
        return cache[insumo_id]
    if member_ids_cache is None:
        member_ids_cache = {}
    member_ids = member_ids_cache.get(insumo_id)
    if member_ids is None:
        canonical_row = next((row for row in canonicalized_active_insumos(limit=5000) if insumo_id in row["member_ids"]), None)
        member_ids = canonical_row["member_ids"] if canonical_row else [insumo_id]
        for member_id in member_ids:
            member_ids_cache[member_id] = member_ids
    latest = (
        CostoInsumo.objects.filter(insumo_id__in=member_ids)
        .order_by("-fecha", "-id")
        .values_list("costo_unitario", flat=True)
        .first()
    )
    cache[insumo_id] = latest
    return latest


def _reprocess_recetas_pending_matching() -> dict[str, int]:
    qs = LineaReceta.objects.filter(~Q(tipo_linea=LineaReceta.TIPO_SUBSECCION)).filter(
        Q(insumo__isnull=True) | Q(match_status=LineaReceta.STATUS_NEEDS_REVIEW) | Q(match_status=LineaReceta.STATUS_REJECTED)
    )
    total = 0
    auto_ok = 0
    review = 0
    rejected = 0
    linked = 0
    cost_cache: dict[int, Decimal | None] = {}
    member_ids_cache: dict[int, list[int]] = {}

    for linea in qs:
        total += 1
        insumo, score, method = match_insumo(linea.insumo_texto or "")
        status = clasificar_match(score)

        linea.match_score = score
        linea.match_method = method
        linea.match_status = status

        if status == LineaReceta.STATUS_REJECTED or not insumo:
            linea.insumo = None
        else:
            linea.insumo = insumo
            linked += 1
            if not linea.unidad_id and insumo.unidad_base_id:
                linea.unidad = insumo.unidad_base
            if (not linea.unidad_texto) and insumo.unidad_base_id and insumo.unidad_base:
                linea.unidad_texto = insumo.unidad_base.codigo
            latest_cost = _latest_cost_by_insumo_cached(insumo.id, cost_cache, member_ids_cache)
            if latest_cost is not None:
                linea.costo_unitario_snapshot = latest_cost

        linea.save(
            update_fields=[
                "insumo",
                "unidad",
                "unidad_texto",
                "costo_unitario_snapshot",
                "match_status",
                "match_method",
                "match_score",
            ]
        )

        if status == LineaReceta.STATUS_AUTO:
            auto_ok += 1
        elif status == LineaReceta.STATUS_NEEDS_REVIEW:
            review += 1
        else:
            rejected += 1

    return {
        "total": total,
        "auto_ok": auto_ok,
        "review": review,
        "rejected": rejected,
        "linked": linked,
    }


@login_required
def importar_archivos(request: HttpRequest) -> HttpResponse:
    if not can_view_inventario(request.user):
        raise PermissionDenied("No tienes permisos para ver Inventario.")

    summary = None
    pendientes_preview = list(request.session.get("inventario_pending_preview", []))[:80]
    warnings: list[str] = []
    drive_info = None
    drive_sync_mode = get_drive_sync_mode()

    if request.method == "POST":
        if not can_manage_inventario(request.user):
            raise PermissionDenied("No tienes permisos para importar archivos de almacén.")

        action = (request.POST.get("action") or "upload").strip().lower()
        if action == "create_alias":
            alias_name = (request.POST.get("alias_name") or "").strip()
            insumo_id = (request.POST.get("insumo_id") or "").strip()
            if not alias_name or not insumo_id:
                messages.error(request, "Debes indicar nombre origen e insumo para crear el alias.")
                return redirect("inventario:importar_archivos")

            insumo = canonical_insumo_by_id(insumo_id)
            if not insumo:
                messages.error(request, "Insumo inválido para crear alias.")
                return redirect("inventario:importar_archivos")

            ok, alias_norm, action_label = _upsert_alias(alias_name, insumo)
            if not ok:
                messages.error(request, "El nombre origen no es válido.")
                return redirect("inventario:importar_archivos")

            _remove_pending_name_from_session(request, alias_norm)
            _remove_pending_name_from_recent_runs(alias_norm)
            messages.success(request, f"Alias {action_label}: '{alias_name}' -> {insumo.nombre}. Ejecuta de nuevo la importación para aplicar el cambio.")
            return redirect("inventario:importar_archivos")

        selected_sources = set(request.POST.getlist("sources")) or set(SOURCE_TO_FILENAME.keys())
        selected_sources = selected_sources.intersection(set(SOURCE_TO_FILENAME.keys()))
        if not selected_sources:
            messages.error(request, "Selecciona al menos una fuente a importar.")
            return redirect("inventario:importar_archivos")

        fuzzy_threshold = int(_to_decimal(request.POST.get("fuzzy_threshold"), "96"))
        create_aliases = bool(request.POST.get("create_aliases"))
        create_missing_insumos = bool(request.POST.get("create_missing_insumos"))
        run_started_at = timezone.now()
        run_source = AlmacenSyncRun.SOURCE_DRIVE if action == "drive" else AlmacenSyncRun.SOURCE_MANUAL

        if action == "drive":
            month_override = (request.POST.get("month") or "").strip() or None
            try:
                drive_result = sync_almacen_from_drive(
                    include_sources=selected_sources,
                    month_override=month_override,
                    fallback_previous=True,
                    fuzzy_threshold=fuzzy_threshold,
                    create_aliases=create_aliases,
                    create_missing_insumos=create_missing_insumos,
                    dry_run=False,
                )
                summary = drive_result.summary
                drive_info = {
                    "folder_name": drive_result.folder_name,
                    "target_month": drive_result.target_month,
                    "used_fallback_month": drive_result.used_fallback_month,
                    "downloaded_sources": drive_result.downloaded_sources,
                }
                if drive_sync_mode.get("forced"):
                    messages.info(
                        request,
                        (
                            "Modo temporal Drive activo: importación bloqueada a "
                            f"{drive_sync_mode.get('forced_month')} · fuentes "
                            f"{', '.join(drive_sync_mode.get('forced_sources', []))}."
                        ),
                    )
                warnings.extend(drive_result.skipped_files)
            except Exception as exc:
                log_sync_run(
                    source=run_source,
                    status=AlmacenSyncRun.STATUS_ERROR,
                    triggered_by=request.user,
                    message=str(exc),
                    started_at=run_started_at,
                )
                messages.error(request, f"Falló la sincronización desde Google Drive: {exc}")
                return redirect("inventario:importar_archivos")
        else:
            uploaded_files = request.FILES.getlist("archivos")
            if not uploaded_files:
                messages.error(request, "Selecciona al menos un archivo .xlsx para importar.")
                return redirect("inventario:importar_archivos")

            saved_sources = set()
            with TemporaryDirectory(prefix="inv-upload-") as tmpdir:
                for f in uploaded_files:
                    source, target_name = _classify_upload(f.name)
                    if not source or not target_name:
                        warnings.append(f"Archivo no reconocido y omitido: {f.name}")
                        continue
                    if source not in selected_sources:
                        warnings.append(f"Archivo omitido por filtro de fuente ({source}): {f.name}")
                        continue
                    _save_uploaded_file(tmpdir, target_name, f)
                    saved_sources.add(source)

                run_sources = selected_sources.intersection(saved_sources)
                if not run_sources:
                    messages.error(request, "Ninguno de los archivos subidos coincide con las fuentes seleccionadas.")
                    return redirect("inventario:importar_archivos")

                try:
                    summary = import_folder(
                        folderpath=tmpdir,
                        include_sources=run_sources,
                        fuzzy_threshold=fuzzy_threshold,
                        create_aliases=create_aliases,
                        create_missing_insumos=create_missing_insumos,
                        dry_run=False,
                    )
                except Exception as exc:
                    log_sync_run(
                        source=run_source,
                        status=AlmacenSyncRun.STATUS_ERROR,
                        triggered_by=request.user,
                        message=str(exc),
                        started_at=run_started_at,
                    )
                    messages.error(request, f"Falló la importación de almacén: {exc}")
                    return redirect("inventario:importar_archivos")

        for w in warnings:
            messages.warning(request, w)

        if summary:
            if drive_info:
                suffix = (
                    f" (Drive: {drive_info['folder_name']} · objetivo {drive_info['target_month']}"
                    f"{' · fallback aplicado' if drive_info['used_fallback_month'] else ''})"
                )
            else:
                suffix = ""
            messages.success(
                request,
                (
                    f"Importación aplicada. Existencias actualizadas: {summary.existencias_updated}, "
                    f"Movimientos creados: {summary.movimientos_created}, "
                    f"Duplicados omitidos: {summary.movimientos_skipped_duplicate}.{suffix}"
                ),
            )
            if summary.unmatched:
                messages.warning(request, f"Quedaron {summary.unmatched} filas sin match.")
            pendientes_preview = summary.pendientes[:80]
            request.session["inventario_pending_preview"] = summary.pendientes[:200]
            log_sync_run(
                source=run_source,
                status=AlmacenSyncRun.STATUS_OK,
                summary=summary,
                triggered_by=request.user,
                folder_name=(drive_info or {}).get("folder_name", ""),
                target_month=(drive_info or {}).get("target_month", ""),
                fallback_used=bool((drive_info or {}).get("used_fallback_month")),
                downloaded_sources=(drive_info or {}).get("downloaded_sources", sorted(selected_sources)),
                message=" | ".join(warnings[:12]),
                started_at=run_started_at,
            )

    context = {
        "pending_grouped": _build_pending_grouped(pendientes_preview),
        "can_manage_inventario": can_manage_inventario(request.user),
        "sources": [
            ("inventario", "Inventario"),
            ("entradas", "Entradas"),
            ("salidas", "Salidas"),
            ("merma", "Merma"),
        ],
        "summary": summary,
        "pendientes_preview": pendientes_preview,
        "expected_names": [
            INVENTARIO_FILE,
            ENTRADAS_FILE,
            SALIDAS_FILE,
            MERMA_FILE,
        ],
        "defaults": {
            "fuzzy_threshold": 96,
            "create_missing_insumos": True,
            "create_aliases": False,
        },
        "current_month": timezone.localdate().strftime("%Y-%m"),
        "drive_info": drive_info,
        "drive_sync_mode": drive_sync_mode,
        "latest_runs": AlmacenSyncRun.objects.select_related("triggered_by").all()[:10],
        "insumo_alias_targets": canonicalized_insumo_selector(limit=800),
    }
    pending_source_buckets: dict[str, dict[str, object]] = {}
    for row in context["pending_grouped"]:
        sources = [item.strip() for item in str(row.get("sources") or "").split(",") if item.strip() and item.strip() != "-"]
        if not sources:
            sources = ["sin fuente"]
        for source in sources:
            bucket = pending_source_buckets.setdefault(
                source,
                {"source": source, "count": 0, "top_name": row.get("nombre_origen") or "", "top_score": 0.0},
            )
            bucket["count"] += int(row.get("count") or 0)
            score_max = float(row.get("score_max") or 0.0)
            if score_max >= float(bucket["top_score"] or 0.0):
                bucket["top_name"] = row.get("nombre_origen") or bucket["top_name"]
                bucket["top_score"] = score_max
    pending_source_cards = sorted(
        pending_source_buckets.values(),
        key=lambda item: (-int(item["count"]), str(item["source"])),
    )
    pending_focus_rows = list(context["pending_grouped"][:3])
    if pending_focus_rows:
        top_pending = pending_focus_rows[0]
        pending_focus = {
            "label": f"{top_pending['count']} filas · {top_pending['sources']}",
            "summary": (
                f"La carga sigue condicionada por '{top_pending['nombre_origen']}' aún sin integrar "
                "contra el catálogo oficial."
            ),
            "action_label": "Abrir catálogo de referencias",
            "action_detail": "Unifica el nombre antes del siguiente sync para evitar errores de inventario y movimiento.",
            "action_url": reverse("inventario:aliases_catalog") + f"?q={quote_plus(top_pending['nombre_origen'])}",
            "tone": "warning",
        }
    else:
        pending_focus = {
            "label": "Homologación al día",
            "summary": "No hay nombres pendientes de integración en la carga actual.",
            "action_label": "Abrir catálogo de referencias",
            "action_detail": "Puedes revisar referencias de forma preventiva si quieres mantener el catálogo limpio.",
            "action_url": reverse("inventario:aliases_catalog"),
            "tone": "success",
        }
    latest_run = context["latest_runs"][0] if context["latest_runs"] else None
    unresolved_count = int((summary.unmatched if summary else 0) or len(pendientes_preview) or 0)
    movement_count = int((summary.movimientos_created if summary else 0) or 0)
    stock_count = int((summary.existencias_updated if summary else 0) or 0)
    latest_status_ok = bool(latest_run and latest_run.status == AlmacenSyncRun.STATUS_OK)
    context["sync_workflow_rows"] = [
        {
            "step": "01",
            "title": "Entrada externa",
            "status": "Ejecutada" if latest_run else "Pendiente",
            "detail": (
                f"Último origen {latest_run.get_source_display()} · {latest_run.started_at:%Y-%m-%d %H:%M}"
                if latest_run
                else "Aún no se registra una carga de almacén en esta sesión operativa."
            ),
        },
        {
            "step": "02",
            "title": "Validación ERP",
            "status": "Abierta" if unresolved_count > 0 else "Cerrada",
            "detail": (
                f"{unresolved_count} artículo(s) todavía requieren cierre contra el catálogo ERP."
                if unresolved_count > 0
                else "Las referencias de la carga actual ya están cerradas contra el catálogo ERP."
            ),
        },
        {
            "step": "03",
            "title": "Aplicación inventario",
            "status": "Aplicada" if stock_count > 0 or movement_count > 0 else "Sin movimiento",
            "detail": f"Existencias {stock_count} · Movimientos {movement_count}.",
        },
        {
            "step": "04",
            "title": "Liberación operativa",
            "status": "Liberada" if latest_status_ok and unresolved_count == 0 else "En curso",
            "detail": (
                "La carga puede alimentar inventario, reabasto y análisis sin bloqueos abiertos."
                if latest_status_ok and unresolved_count == 0
                else "Todavía hay trabajo operativo antes de liberar completamente la carga."
            ),
        },
    ]
    context["sync_release_summary"] = {
        "title": "Carga liberada" if latest_status_ok and unresolved_count == 0 else "Carga en curso",
        "detail": (
            "La carga actual ya quedó lista para operar en inventario y reabasto."
            if latest_status_ok and unresolved_count == 0
            else "Revisa las referencias abiertas antes de considerar cerrada esta carga."
        ),
        "tone": "success" if latest_status_ok and unresolved_count == 0 else "warning",
    }
    context["pending_source_cards"] = pending_source_cards
    context["pending_focus_rows"] = pending_focus_rows
    context["pending_focus"] = pending_focus
    context["erp_governance_rows"] = [
        {
            "front": "Carga externa",
            "owner": "Inventario / Operación",
            "blockers": unresolved_count,
            "completion": 100 if latest_run and unresolved_count == 0 else (65 if latest_run else 25),
            "detail": (
                "Última corrida aplicada y lista para operar."
                if latest_run and unresolved_count == 0
                else "La carga todavía depende del cierre de referencias ERP."
            ),
            "next_step": (
                "Mantener monitoreo preventivo del siguiente corte."
                if latest_run and unresolved_count == 0
                else "Resolver referencias abiertas antes del siguiente sync."
            ),
            "url": reverse("inventario:importar_archivos"),
            "cta": "Revisar carga",
        },
        {
            "front": "Referencias ERP",
            "owner": "Maestros / Inventario",
            "blockers": unresolved_count,
            "completion": 100 if unresolved_count == 0 else 70,
            "detail": (
                "Las referencias de la carga actual ya quedaron integradas al maestro."
                if unresolved_count == 0
                else f"{unresolved_count} referencia(s) siguen abiertas contra el maestro ERP."
            ),
            "next_step": (
                "Mantener el catálogo estable."
                if unresolved_count == 0
                else "Abrir referencias ERP y cerrar pendientes del catálogo."
            ),
            "url": reverse("inventario:aliases_catalog"),
            "cta": "Abrir referencias ERP",
        },
        {
            "front": "Inventario operativo",
            "owner": "Inventario / Reabasto",
            "blockers": 0 if stock_count > 0 or movement_count > 0 else unresolved_count,
            "completion": 100 if stock_count > 0 or movement_count > 0 else (80 if unresolved_count == 0 else 55),
            "detail": (
                f"Existencias {stock_count} · Movimientos {movement_count} ya aplicados."
                if stock_count > 0 or movement_count > 0
                else "La aplicación a inventario todavía no cierra completamente."
            ),
            "next_step": (
                "Continuar con monitoreo y alertas."
                if stock_count > 0 or movement_count > 0
                else "Validar existencias y movimientos posteriores a la carga."
            ),
            "url": reverse("inventario:existencias"),
            "cta": "Abrir inventario",
        },
    ]
    context["executive_radar_rows"] = _inventario_executive_radar_rows(context["erp_governance_rows"])
    return render(request, "inventario/importar_archivos.html", context)


MASTER_NORMALIZE_SCOPES = ("all", "insumos", "recetas", "aliases_insumo", "receta_codigos_point")
MASTER_DUPLICATES_SCOPES = ("all", "insumos", "recetas", "proveedores", "codigos_point")


def _read_master_normalize_filters(params) -> tuple[str, str, int, int]:
    scope = str(params.get("master_scope") or "all").strip().lower()
    if scope not in MASTER_NORMALIZE_SCOPES:
        scope = "all"
    q = (params.get("master_q") or "").strip()
    limit = int(_to_decimal(params.get("master_limit"), "80"))
    limit = max(1, min(5000, limit))
    offset = int(_to_decimal(params.get("master_offset"), "0"))
    offset = max(0, min(100000, offset))
    return scope, q, limit, offset


def _read_master_duplicates_filters(params) -> tuple[str, str, bool, int, int, int]:
    scope = str(params.get("master_dup_scope") or "all").strip().lower()
    if scope not in MASTER_DUPLICATES_SCOPES:
        scope = "all"
    q = (params.get("master_dup_q") or "").strip()
    include_inactive = str(params.get("master_dup_include_inactive") or "").strip() in {"1", "true", "on", "yes"}
    min_count = int(_to_decimal(params.get("master_dup_min_count"), "2"))
    min_count = max(2, min(200, min_count))
    limit = int(_to_decimal(params.get("master_dup_limit"), "120"))
    limit = max(1, min(3000, limit))
    offset = int(_to_decimal(params.get("master_dup_offset"), "0"))
    offset = max(0, min(100000, offset))
    return scope, q, include_inactive, min_count, limit, offset


def _master_query_by_scope(scope: str, q: str):
    q_norm = normalizar_nombre(q) if q else ""
    if scope == "insumos":
        qs = Insumo.objects.all().order_by("id")
        if q:
            qs = qs.filter(
                Q(nombre__icontains=q)
                | Q(nombre_normalizado__icontains=q_norm)
                | Q(codigo__icontains=q)
                | Q(codigo_point__icontains=q)
            )
        return qs
    if scope == "recetas":
        qs = Receta.objects.all().order_by("id")
        if q:
            qs = qs.filter(
                Q(nombre__icontains=q)
                | Q(nombre_normalizado__icontains=q_norm)
                | Q(codigo_point__icontains=q)
            )
        return qs
    if scope == "aliases_insumo":
        qs = InsumoAlias.objects.select_related("insumo").all().order_by("id")
        if q:
            qs = qs.filter(
                Q(nombre__icontains=q)
                | Q(nombre_normalizado__icontains=q_norm)
                | Q(insumo__nombre__icontains=q)
            )
        return qs
    if scope == "receta_codigos_point":
        qs = RecetaCodigoPointAlias.objects.select_related("receta").all().order_by("id")
        if q:
            qs = qs.filter(
                Q(codigo_point__icontains=q)
                | Q(codigo_point_normalizado__icontains=normalizar_codigo_point(q))
                | Q(nombre_point__icontains=q)
                | Q(receta__nombre__icontains=q)
            )
        return qs
    return None


def _master_serialize_normalize_row(scope: str, obj, current: str, suggested: str) -> dict[str, object]:
    if scope == "insumos":
        return {
            "scope": scope,
            "model": "maestros.Insumo",
            "id": obj.id,
            "label": obj.nombre,
            "actual": current,
            "sugerido": suggested,
            "changed": bool((current or "") != (suggested or "")),
            "codigo": obj.codigo or "",
            "codigo_point": obj.codigo_point or "",
        }
    if scope == "recetas":
        return {
            "scope": scope,
            "model": "recetas.Receta",
            "id": obj.id,
            "label": obj.nombre,
            "actual": current,
            "sugerido": suggested,
            "changed": bool((current or "") != (suggested or "")),
            "codigo_point": obj.codigo_point or "",
        }
    if scope == "aliases_insumo":
        return {
            "scope": scope,
            "model": "maestros.InsumoAlias",
            "id": obj.id,
            "label": obj.nombre,
            "actual": current,
            "sugerido": suggested,
            "changed": bool((current or "") != (suggested or "")),
            "insumo_id": obj.insumo_id,
            "insumo": obj.insumo.nombre if obj.insumo_id else "",
        }
    return {
        "scope": scope,
        "model": "recetas.RecetaCodigoPointAlias",
        "id": obj.id,
        "label": obj.codigo_point or "",
        "actual": current,
        "sugerido": suggested,
        "changed": bool((current or "") != (suggested or "")),
        "receta_id": obj.receta_id,
        "receta": obj.receta.nombre if obj.receta_id else "",
    }


def _run_master_normalize(scope: str, q: str, limit: int, offset: int, apply_changes: bool = False) -> dict[str, object]:
    selected_scopes = (
        ["insumos", "recetas", "aliases_insumo", "receta_codigos_point"] if scope == "all" else [scope]
    )
    by_scope: dict[str, dict[str, int]] = {}
    items: list[dict[str, object]] = []
    total_candidates = 0
    total_changed = 0
    total_updated = 0

    for current_scope in selected_scopes:
        qs = _master_query_by_scope(current_scope, q)
        if qs is None:
            continue
        scoped_total = qs.count()
        scoped_rows = qs[offset : offset + limit]
        scoped_changed = 0
        scoped_updated = 0

        for obj in scoped_rows:
            if current_scope in {"insumos", "recetas", "aliases_insumo"}:
                current_val = str(getattr(obj, "nombre_normalizado", "") or "")
                suggested = normalizar_nombre(getattr(obj, "nombre", "") or "")
                field_name = "nombre_normalizado"
            else:
                current_val = str(getattr(obj, "codigo_point_normalizado", "") or "")
                suggested = normalizar_codigo_point(getattr(obj, "codigo_point", "") or "")
                field_name = "codigo_point_normalizado"

            changed = current_val != suggested
            if changed:
                scoped_changed += 1
                if apply_changes:
                    setattr(obj, field_name, suggested)
                    obj.save(update_fields=[field_name])
                    scoped_updated += 1

            items.append(_master_serialize_normalize_row(current_scope, obj, current_val, suggested))

        total_candidates += scoped_total
        total_changed += scoped_changed
        total_updated += scoped_updated
        by_scope[current_scope] = {
            "candidates": scoped_total,
            "returned": len(scoped_rows),
            "changed": scoped_changed,
            "updated": scoped_updated,
        }

    return {
        "filters": {
            "scope": scope,
            "q": q,
            "limit": limit,
            "offset": offset,
            "scopes_evaluated": selected_scopes,
            "apply_changes": bool(apply_changes),
        },
        "totales": {
            "candidates": total_candidates,
            "changed": total_changed,
            "updated": total_updated,
            "returned": len(items),
        },
        "by_scope": by_scope,
        "items": items,
    }


def _build_master_duplicate_group(group_type: str, key: str, members: list[dict[str, object]]) -> dict[str, object]:
    members_sorted = sorted(
        members,
        key=lambda item: (
            str(item.get("nombre") or item.get("label") or "").lower(),
            int(item.get("id") or 0),
        ),
    )
    return {
        "group_type": group_type,
        "duplicate_key": key,
        "count": len(members_sorted),
        "members": members_sorted,
    }


def _build_master_duplicates(
    scope: str,
    q: str,
    include_inactive: bool,
    min_count: int,
    limit: int,
    offset: int,
) -> dict[str, object]:
    q_norm = normalizar_nombre(q) if q else ""
    selected_scopes = ["insumos", "recetas", "proveedores", "codigos_point"] if scope == "all" else [scope]
    groups: list[dict[str, object]] = []
    by_scope_totals = {k: 0 for k in selected_scopes}

    if "insumos" in selected_scopes:
        qs = Insumo.objects.all().order_by("id")
        if not include_inactive:
            qs = qs.filter(activo=True)
        if q:
            qs = qs.filter(
                Q(nombre__icontains=q)
                | Q(nombre_normalizado__icontains=q_norm)
                | Q(codigo__icontains=q)
                | Q(codigo_point__icontains=q)
            )
        grouped = (
            qs.values("nombre_normalizado")
            .annotate(total=Count("id"))
            .filter(total__gte=min_count)
            .order_by("-total", "nombre_normalizado")
        )
        for row in grouped:
            key = str(row.get("nombre_normalizado") or "")
            members = [
                {
                    "model": "maestros.Insumo",
                    "id": int(obj.id),
                    "nombre": obj.nombre,
                    "activo": bool(obj.activo),
                    "codigo_point": obj.codigo_point or "",
                }
                for obj in qs.filter(nombre_normalizado=key).order_by("id")
            ]
            groups.append(_build_master_duplicate_group("insumos", key, members))
        by_scope_totals["insumos"] = len([g for g in groups if g["group_type"] == "insumos"])

    if "recetas" in selected_scopes:
        qs = Receta.objects.all().order_by("id")
        if q:
            qs = qs.filter(
                Q(nombre__icontains=q)
                | Q(nombre_normalizado__icontains=q_norm)
                | Q(codigo_point__icontains=q)
            )
        grouped = (
            qs.values("nombre_normalizado")
            .annotate(total=Count("id"))
            .filter(total__gte=min_count)
            .order_by("-total", "nombre_normalizado")
        )
        for row in grouped:
            key = str(row.get("nombre_normalizado") or "")
            members = [
                {
                    "model": "recetas.Receta",
                    "id": int(obj.id),
                    "nombre": obj.nombre,
                    "activo": True,
                    "codigo_point": obj.codigo_point or "",
                }
                for obj in qs.filter(nombre_normalizado=key).order_by("id")
            ]
            groups.append(_build_master_duplicate_group("recetas", key, members))
        by_scope_totals["recetas"] = len([g for g in groups if g["group_type"] == "recetas"])

    if "proveedores" in selected_scopes:
        qs = Proveedor.objects.all().order_by("id")
        if not include_inactive:
            qs = qs.filter(activo=True)
        if q:
            qs = qs.filter(nombre__icontains=q)
        by_key: dict[str, list[dict[str, object]]] = defaultdict(list)
        for provider in qs:
            key = normalizar_nombre(provider.nombre or "")
            if not key:
                continue
            by_key[key].append(
                {
                    "model": "maestros.Proveedor",
                    "id": int(provider.id),
                    "nombre": provider.nombre,
                    "activo": bool(provider.activo),
                    "codigo_point": "",
                }
            )
        for key, members in sorted(by_key.items(), key=lambda item: (-len(item[1]), item[0])):
            if len(members) < min_count:
                continue
            groups.append(_build_master_duplicate_group("proveedores", key, members))
        by_scope_totals["proveedores"] = len([g for g in groups if g["group_type"] == "proveedores"])

    if "codigos_point" in selected_scopes:
        by_code: dict[str, list[dict[str, object]]] = defaultdict(list)
        insumos_qs = Insumo.objects.exclude(codigo_point="").order_by("id")
        recetas_qs = Receta.objects.exclude(codigo_point="").order_by("id")
        if not include_inactive:
            insumos_qs = insumos_qs.filter(activo=True)
        if q:
            insumos_qs = insumos_qs.filter(Q(codigo_point__icontains=q) | Q(nombre__icontains=q))
            recetas_qs = recetas_qs.filter(Q(codigo_point__icontains=q) | Q(nombre__icontains=q))
        for insumo in insumos_qs:
            key = normalizar_codigo_point(insumo.codigo_point or "")
            if not key:
                continue
            by_code[key].append(
                {
                    "model": "maestros.Insumo",
                    "id": int(insumo.id),
                    "nombre": insumo.nombre,
                    "activo": bool(insumo.activo),
                    "codigo_point": insumo.codigo_point or "",
                }
            )
        for receta in recetas_qs:
            key = normalizar_codigo_point(receta.codigo_point or "")
            if not key:
                continue
            by_code[key].append(
                {
                    "model": "recetas.Receta",
                    "id": int(receta.id),
                    "nombre": receta.nombre,
                    "activo": True,
                    "codigo_point": receta.codigo_point or "",
                }
            )
        for key, members in sorted(by_code.items(), key=lambda item: (-len(item[1]), item[0])):
            if len(members) < min_count:
                continue
            groups.append(_build_master_duplicate_group("codigos_point", key, members))
        by_scope_totals["codigos_point"] = len([g for g in groups if g["group_type"] == "codigos_point"])

    groups.sort(key=lambda row: (-int(row.get("count") or 0), str(row.get("duplicate_key") or "").lower()))
    total_groups = len(groups)
    page_groups = groups[offset : offset + limit]

    return {
        "filters": {
            "scope": scope,
            "q": q,
            "include_inactive": include_inactive,
            "min_count": min_count,
            "limit": limit,
            "offset": offset,
        },
        "totales": {
            "groups_total": total_groups,
            "groups_returned": len(page_groups),
            "by_scope": by_scope_totals,
        },
        "items": page_groups,
    }


def _export_master_duplicates_csv(groups: list[dict[str, object]]) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="master_duplicates.csv"'
    writer = csv.writer(response)
    writer.writerow(["group_type", "duplicate_key", "count", "model", "id", "nombre", "activo", "codigo_point"])
    for group in groups:
        for member in group.get("members") or []:
            writer.writerow(
                [
                    group.get("group_type") or "",
                    group.get("duplicate_key") or "",
                    group.get("count") or 0,
                    member.get("model") or "",
                    member.get("id") or "",
                    member.get("nombre") or "",
                    "1" if member.get("activo") else "0",
                    member.get("codigo_point") or "",
                ]
            )
    return response


def _export_master_duplicates_xlsx(groups: list[dict[str, object]]) -> HttpResponse:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "duplicates"
    sheet.append(["group_type", "duplicate_key", "count", "model", "id", "nombre", "activo", "codigo_point"])
    for group in groups:
        for member in group.get("members") or []:
            sheet.append(
                [
                    group.get("group_type") or "",
                    group.get("duplicate_key") or "",
                    group.get("count") or 0,
                    member.get("model") or "",
                    member.get("id") or "",
                    member.get("nombre") or "",
                    "1" if member.get("activo") else "0",
                    member.get("codigo_point") or "",
                ]
            )
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="master_duplicates.xlsx"'
    return response


def _merge_insumo_into_target(source: Insumo, target: Insumo) -> dict[str, int]:
    if source.id == target.id:
        return {
            "aliases_updated": 0,
            "costos_updated": 0,
            "lineas_updated": 0,
            "movimientos_updated": 0,
            "ajustes_updated": 0,
            "solicitudes_updated": 0,
            "existencia_merged": 0,
        }

    alias_count = InsumoAlias.objects.filter(insumo_id=source.id).update(insumo_id=target.id)
    costo_count = CostoInsumo.objects.filter(insumo_id=source.id).update(insumo_id=target.id)
    linea_count = LineaReceta.objects.filter(insumo_id=source.id).update(insumo_id=target.id)
    movimiento_count = MovimientoInventario.objects.filter(insumo_id=source.id).update(insumo_id=target.id)
    ajuste_count = AjusteInventario.objects.filter(insumo_id=source.id).update(insumo_id=target.id)
    solicitud_count = SolicitudCompra.objects.filter(insumo_id=source.id).update(insumo_id=target.id)

    existencia_merged = 0
    source_ex = ExistenciaInsumo.objects.filter(insumo_id=source.id).first()
    if source_ex:
        target_ex = ExistenciaInsumo.objects.filter(insumo_id=target.id).first()
        if not target_ex:
            source_ex.insumo_id = target.id
            source_ex.actualizado_en = timezone.now()
            source_ex.save(update_fields=["insumo", "actualizado_en"])
            existencia_merged = 1
        else:
            target_ex.stock_actual = _to_decimal(target_ex.stock_actual) + _to_decimal(source_ex.stock_actual)
            target_ex.punto_reorden = max(_to_decimal(target_ex.punto_reorden), _to_decimal(source_ex.punto_reorden))
            target_ex.stock_minimo = max(_to_decimal(target_ex.stock_minimo), _to_decimal(source_ex.stock_minimo))
            target_ex.stock_maximo = max(_to_decimal(target_ex.stock_maximo), _to_decimal(source_ex.stock_maximo))
            target_ex.inventario_promedio = max(
                _to_decimal(target_ex.inventario_promedio),
                _to_decimal(source_ex.inventario_promedio),
            )
            target_ex.dias_llegada_pedido = max(
                int(target_ex.dias_llegada_pedido or 0),
                int(source_ex.dias_llegada_pedido or 0),
            )
            target_ex.consumo_diario_promedio = max(
                _to_decimal(target_ex.consumo_diario_promedio),
                _to_decimal(source_ex.consumo_diario_promedio),
            )
            target_ex.actualizado_en = timezone.now()
            target_ex.save(
                update_fields=[
                    "stock_actual",
                    "punto_reorden",
                    "stock_minimo",
                    "stock_maximo",
                    "inventario_promedio",
                    "dias_llegada_pedido",
                    "consumo_diario_promedio",
                    "actualizado_en",
                ]
            )
            source_ex.delete()
            existencia_merged = 1

    if source.proveedor_principal_id and not target.proveedor_principal_id:
        target.proveedor_principal_id = source.proveedor_principal_id
    if source.unidad_base_id and not target.unidad_base_id:
        target.unidad_base_id = source.unidad_base_id
    if source.codigo_point and not target.codigo_point:
        target.codigo_point = source.codigo_point
    if source.nombre_point and not target.nombre_point:
        target.nombre_point = source.nombre_point
    target.save(
        update_fields=[
            "proveedor_principal",
            "unidad_base",
            "codigo_point",
            "nombre_point",
        ]
    )

    source.activo = False
    source.save(update_fields=["activo"])

    return {
        "aliases_updated": alias_count,
        "costos_updated": costo_count,
        "lineas_updated": linea_count,
        "movimientos_updated": movimiento_count,
        "ajustes_updated": ajuste_count,
        "solicitudes_updated": solicitud_count,
        "existencia_merged": existencia_merged,
    }


@login_required
def aliases_catalog(request: HttpRequest) -> HttpResponse:
    if not can_view_inventario(request.user):
        raise PermissionDenied("No tienes permisos para ver Inventario.")

    if request.method == "POST":
        if not can_manage_inventario(request.user):
            raise PermissionDenied("No tienes permisos para administrar referencias de catálogo.")

        action = (request.POST.get("action") or "").strip().lower()
        next_q = (request.POST.get("next_q") or "").strip()

        if action == "create":
            alias_name = (request.POST.get("alias_name") or "").strip()
            insumo_id = (request.POST.get("insumo_id") or "").strip()
            if not alias_name or not insumo_id:
                messages.error(request, "Debes indicar nombre origen e insumo.")
            else:
                insumo = canonical_insumo_by_id(insumo_id)
                if not insumo:
                    messages.error(request, "Insumo inválido.")
                else:
                    ok, alias_norm, action_label = _upsert_alias(alias_name, insumo)
                    if ok:
                        _remove_pending_name_from_session(request, alias_norm)
                        _remove_pending_name_from_recent_runs(alias_norm)
                        messages.success(request, f"Alias {action_label}: '{alias_name}' -> {insumo.nombre}.")
                        point_resolved, recetas_resolved = _resolve_cross_source_with_alias(alias_name, insumo)
                        if point_resolved or recetas_resolved:
                            messages.info(
                                request,
                                (
                                    "Integración cruzada aplicada: "
                                    f"Comerciales resueltos {point_resolved}, "
                                    f"líneas recetas resueltas {recetas_resolved}."
                                ),
                            )
                    else:
                        messages.error(request, "El nombre origen no es válido.")

        elif action == "apply_suggestion":
            alias_name = (request.POST.get("alias_name") or "").strip()
            suggestion = (request.POST.get("suggestion") or "").strip()
            min_score = float(_to_decimal(request.POST.get("score_min"), "90"))
            min_score = max(0.0, min(100.0, min_score))
            if not alias_name or not suggestion:
                messages.error(request, "Faltan datos para aplicar sugerencia.")
            else:
                insumo = (
                    Insumo.objects.filter(activo=True, nombre_normalizado=normalizar_nombre(suggestion))
                    .only("id", "nombre")
                    .first()
                )
                if not insumo:
                    candidate, candidate_score, _ = match_insumo(suggestion)
                    if candidate and float(candidate_score or 0.0) >= min_score:
                        insumo = candidate
                if not insumo:
                    messages.error(request, f"No se pudo resolver '{suggestion}' como insumo activo.")
                else:
                    ok, alias_norm, action_label = _upsert_alias(alias_name, insumo)
                    if ok:
                        _remove_pending_name_from_session(request, alias_norm)
                        _remove_pending_name_from_recent_runs(alias_norm)
                        point_resolved, recetas_resolved = _resolve_cross_source_with_alias(alias_name, insumo)
                        messages.success(
                            request,
                            (
                                f"Sugerencia aplicada ({action_label}): '{alias_name}' -> {insumo.nombre}. "
                                f"Comerciales resueltos: {point_resolved}. Recetas resueltas: {recetas_resolved}."
                            ),
                        )
                    else:
                        messages.error(request, "No se pudo aplicar la sugerencia por normalización inválida.")

        elif action == "import_bulk":
            archivo = request.FILES.get("archivo_aliases")
            min_score = float(_to_decimal(request.POST.get("score_min"), "90"))
            min_score = max(0.0, min(100.0, min_score))

            if not archivo:
                messages.error(request, "Debes seleccionar un archivo .csv o .xlsx para importar referencias.")
            else:
                try:
                    rows = _read_alias_import_rows(archivo)
                except ValueError as exc:
                    messages.error(request, str(exc))
                    rows = []
                except Exception:
                    messages.error(request, "No se pudo leer el archivo de referencias. Verifica formato y columnas.")
                    rows = []

                if rows:
                    insumo_exact_map = {
                        i.nombre_normalizado: i
                        for i in Insumo.objects.filter(activo=True).only("id", "nombre", "nombre_normalizado")
                    }
                    match_cache: dict[str, tuple[Insumo | None, float, str]] = {}
                    created = 0
                    updated = 0
                    invalid = 0
                    unresolved = 0
                    cleaned_norms: set[str] = set()
                    unresolved_preview: list[dict] = []
                    cross_resolve_targets: dict[str, tuple[str, Insumo]] = {}

                    for idx, row in enumerate(rows, start=2):
                        alias_name = str(row.get("alias") or "").strip()
                        insumo_raw = str(row.get("insumo") or "").strip()
                        reason = ""
                        resolved_insumo = None
                        score = 0.0
                        method = "-"
                        suggested_name = ""

                        if not alias_name:
                            reason = "Alias vacío."
                            invalid += 1
                        else:
                            if insumo_raw:
                                resolved_insumo = insumo_exact_map.get(normalizar_nombre(insumo_raw))
                                if resolved_insumo:
                                    score = 100.0
                                    method = "EXACT_NAME"
                                    suggested_name = resolved_insumo.nombre
                                else:
                                    lookup_key = normalizar_nombre(insumo_raw)
                                    if lookup_key in match_cache:
                                        candidate, candidate_score, candidate_method = match_cache[lookup_key]
                                    else:
                                        candidate, candidate_score, candidate_method = match_insumo(insumo_raw)
                                        match_cache[lookup_key] = (candidate, candidate_score, candidate_method)
                                    if candidate and candidate_score >= min_score:
                                        resolved_insumo = candidate
                                        score = float(candidate_score or 0.0)
                                        method = candidate_method or "FUZZY"
                                        suggested_name = candidate.nombre
                                    else:
                                        suggested_name = candidate.nombre if candidate else ""
                                        score = float(candidate_score or 0.0) if candidate else 0.0
                                        method = candidate_method or "NO_MATCH"
                                        reason = f"Insumo no resuelto (score<{min_score:.1f})."
                            else:
                                lookup_key = normalizar_nombre(alias_name)
                                if lookup_key in match_cache:
                                    candidate, candidate_score, candidate_method = match_cache[lookup_key]
                                else:
                                    candidate, candidate_score, candidate_method = match_insumo(alias_name)
                                    match_cache[lookup_key] = (candidate, candidate_score, candidate_method)
                                if candidate and candidate_score >= min_score:
                                    resolved_insumo = candidate
                                    score = float(candidate_score or 0.0)
                                    method = candidate_method or "FUZZY"
                                    suggested_name = candidate.nombre
                                else:
                                    suggested_name = candidate.nombre if candidate else ""
                                    score = float(candidate_score or 0.0) if candidate else 0.0
                                    method = candidate_method or "NO_MATCH"
                                    reason = "Sin columna 'insumo' resoluble para esta fila."

                        if not reason and resolved_insumo:
                            ok, alias_norm, action_label = _upsert_alias(alias_name, resolved_insumo)
                            if ok:
                                if action_label == "creado":
                                    created += 1
                                else:
                                    updated += 1
                                cleaned_norms.add(alias_norm)
                                cross_resolve_targets[alias_norm] = (alias_name, resolved_insumo)
                            else:
                                reason = "Alias inválido tras normalización."
                                invalid += 1

                        if reason:
                            unresolved += 1
                            if len(unresolved_preview) < 200:
                                unresolved_preview.append(
                                    {
                                        "row": idx,
                                        "alias": alias_name,
                                        "insumo_archivo": insumo_raw,
                                        "sugerencia": suggested_name,
                                        "score": score,
                                        "method": method,
                                        "motivo": reason,
                                    }
                                )

                    if cleaned_norms:
                        _remove_pending_names_from_session(request, cleaned_norms)
                        _remove_pending_names_from_recent_runs(cleaned_norms)

                    point_resolved_total = 0
                    recetas_resolved_total = 0
                    for alias_name, resolved_insumo in cross_resolve_targets.values():
                        point_resolved, recetas_resolved = _resolve_cross_source_with_alias(alias_name, resolved_insumo)
                        point_resolved_total += point_resolved
                        recetas_resolved_total += recetas_resolved

                    request.session["inventario_alias_import_preview"] = unresolved_preview
                    request.session["inventario_alias_import_stats"] = {
                        "file_name": archivo.name,
                        "rows_total": len(rows),
                        "created": created,
                        "updated": updated,
                        "invalid": invalid,
                        "unresolved": unresolved,
                        "point_resolved": point_resolved_total,
                        "recetas_resolved": recetas_resolved_total,
                        "score_min": min_score,
                    }

                    messages.success(
                        request,
                        (
                            "Importación masiva de referencias completada. "
                            f"Filas: {len(rows)}. Creados: {created}. Actualizados: {updated}. "
                            f"Comerciales resueltos: {point_resolved_total}. Recetas resueltas: {recetas_resolved_total}."
                        ),
                    )
                    if unresolved:
                        messages.warning(
                            request,
                            (
                                f"Quedaron {unresolved} filas sin resolver o inválidas. "
                                "Revísalas en el bloque 'Pendientes de importación'."
                            ),
                        )
                else:
                    messages.warning(request, "El archivo no contiene filas para importar.")

        elif action == "bulk_reassign":
            insumo_id = (request.POST.get("insumo_id") or "").strip()
            alias_ids = [a for a in request.POST.getlist("alias_ids") if a.isdigit()]
            if not insumo_id or not alias_ids:
                messages.error(request, "Selecciona referencias e insumo destino para la reasignación masiva.")
            else:
                insumo = canonical_insumo_by_id(insumo_id)
                if not insumo:
                    messages.error(request, "Insumo destino inválido.")
                else:
                    aliases_to_update = list(
                        InsumoAlias.objects.filter(id__in=alias_ids).exclude(insumo=insumo).only("id", "nombre")
                    )
                    updated = 0
                    point_resolved_total = 0
                    recetas_resolved_total = 0
                    cleaned_norms: set[str] = set()

                    for alias in aliases_to_update:
                        alias.insumo = insumo
                        alias.save(update_fields=["insumo"])
                        updated += 1
                        cleaned_norms.add(normalizar_nombre(alias.nombre))
                        point_resolved, recetas_resolved = _resolve_cross_source_with_alias(alias.nombre, insumo)
                        point_resolved_total += point_resolved
                        recetas_resolved_total += recetas_resolved

                    if cleaned_norms:
                        _remove_pending_names_from_session(request, cleaned_norms)
                        _remove_pending_names_from_recent_runs(cleaned_norms)

                    messages.success(
                        request,
                        (
                            f"Referencias reasignadas a {insumo.nombre}: {updated}. "
                            f"Comerciales resueltos: {point_resolved_total}. "
                            f"Recetas resueltas: {recetas_resolved_total}."
                        ),
                    )

        elif action == "delete":
            alias_id = (request.POST.get("alias_id") or "").strip()
            alias = InsumoAlias.objects.filter(pk=alias_id).first()
            if not alias:
                messages.error(request, "Alias no encontrado.")
            else:
                alias_display = alias.nombre
                alias.delete()
                messages.success(request, f"Alias eliminado: {alias_display}.")

        elif action == "clear_pending":
            request.session["inventario_pending_preview"] = []
            hide_run_id = (request.POST.get("hide_run_id") or "").strip()
            if hide_run_id.isdigit():
                request.session["inventario_hidden_pending_run_id"] = int(hide_run_id)
            messages.success(request, "Pendientes en pantalla limpiados.")
        elif action == "reset_hidden_pending":
            request.session.pop("inventario_hidden_pending_run_id", None)
            messages.success(request, "Se restauró la visibilidad de pendientes recientes.")
        elif action == "load_pending_run":
            run_id_raw = (request.POST.get("run_id") or "").strip()
            run = AlmacenSyncRun.objects.filter(pk=run_id_raw).first() if run_id_raw.isdigit() else None
            if not run:
                messages.error(request, "Run de sincronización no encontrado.")
            elif not isinstance(run.pending_preview, list) or not run.pending_preview:
                messages.error(request, "Ese run no tiene pendientes guardados para mostrar.")
            else:
                request.session["inventario_pending_preview"] = list(run.pending_preview)[:200]
                request.session.pop("inventario_hidden_pending_run_id", None)
                messages.success(
                    request,
                    f"Pendientes cargados desde run {run.id} ({run.started_at:%Y-%m-%d %H:%M}).",
                )
        elif action == "auto_apply_suggestions":
            min_score = float(_to_decimal(request.POST.get("auto_min_score"), "90"))
            min_score = max(0.0, min(100.0, min_score))
            max_rows = int(_to_decimal(request.POST.get("auto_max_rows"), "80"))
            max_rows = max(1, min(500, max_rows))
            min_sources = int(_to_decimal(request.POST.get("auto_min_sources"), "2"))
            min_sources = max(1, min(3, min_sources))
            _, cross_q_norm_post, cross_only_suggested_post, cross_min_sources_post, cross_score_min_post = (
                _read_cross_filters(request.POST)
            )
            cross_point_tipo_post, point_tipos_filter_post = _read_cross_point_tipo(request.POST)
            cross_source_post = _read_cross_source(request.POST)
            cross_limit_post, cross_offset_post, cross_sort_by_post, cross_sort_dir_post = _read_cross_table_controls(
                request.POST
            )

            pending_preview = _load_visible_pending_preview(request, max_rows=500, max_runs=20)
            pending_grouped = _build_pending_grouped(pending_preview)
            cross_unified_rows, _, _ = _build_cross_unified_rows(
                pending_grouped,
                point_tipos=point_tipos_filter_post,
            )
            cross_unified_rows = _apply_cross_filters(
                cross_unified_rows,
                cross_q_norm=cross_q_norm_post,
                cross_only_suggested=cross_only_suggested_post,
                cross_min_sources=cross_min_sources_post,
                cross_score_min=cross_score_min_post,
            )
            if cross_source_post == "ALMACEN":
                cross_unified_rows = [row for row in cross_unified_rows if int(row.get("almacen_count") or 0) > 0]
            elif cross_source_post == "POINT":
                cross_unified_rows = [row for row in cross_unified_rows if int(row.get("point_count") or 0) > 0]
            elif cross_source_post == "RECETAS":
                cross_unified_rows = [row for row in cross_unified_rows if int(row.get("receta_count") or 0) > 0]
            cross_unified_rows = _sort_cross_rows(
                cross_unified_rows,
                sort_by=cross_sort_by_post,
                sort_dir=cross_sort_dir_post,
            )
            cross_unified_rows = cross_unified_rows[cross_offset_post : cross_offset_post + cross_limit_post]

            created = 0
            updated = 0
            skipped_no_suggestion = 0
            skipped_low_score = 0
            skipped_low_sources = 0
            skipped_unresolved = 0
            skipped_invalid = 0
            point_resolved_total = 0
            recetas_resolved_total = 0
            processed = 0
            if not cross_unified_rows:
                messages.info(request, "No hay pendientes visibles para auto-aplicar sugerencias.")
            else:
                insumo_map = {
                    i.nombre_normalizado: i
                    for i in Insumo.objects.filter(activo=True).only("id", "nombre", "nombre_normalizado")
                }
                cleaned_norms: set[str] = set()

                for row in cross_unified_rows:
                    if processed >= max_rows:
                        break

                    if int(row.get("sources_active") or 0) < min_sources:
                        skipped_low_sources += 1
                        continue

                    suggestion = str(row.get("suggestion") or "").strip()
                    if not suggestion:
                        skipped_no_suggestion += 1
                        continue

                    score_max = float(row.get("score_max") or 0.0)
                    if score_max < min_score:
                        skipped_low_score += 1
                        continue

                    insumo = insumo_map.get(normalizar_nombre(suggestion))
                    if not insumo:
                        candidate, candidate_score, _ = match_insumo(suggestion)
                        if candidate and float(candidate_score or 0.0) >= min_score:
                            insumo = candidate
                    if not insumo:
                        skipped_unresolved += 1
                        continue

                    ok, alias_norm, action_label = _upsert_alias(row["nombre_muestra"], insumo)
                    if not ok:
                        skipped_invalid += 1
                        continue

                    processed += 1
                    if action_label == "creado":
                        created += 1
                    else:
                        updated += 1
                    cleaned_norms.add(alias_norm)

                    point_resolved, recetas_resolved = _resolve_cross_source_with_alias(row["nombre_muestra"], insumo)
                    point_resolved_total += point_resolved
                    recetas_resolved_total += recetas_resolved

                if cleaned_norms:
                    _remove_pending_names_from_session(request, cleaned_norms)
                    _remove_pending_names_from_recent_runs(cleaned_norms)

                messages.success(
                    request,
                    (
                        "Auto-aplicación completada. "
                        f"Creados: {created}, actualizados: {updated}, "
                        f"Comerciales resueltos: {point_resolved_total}, "
                        f"Recetas resueltas: {recetas_resolved_total}."
                    ),
                )
                messages.info(
                    request,
                    (
                        "Omitidos → "
                        f"fuentes<{min_sources}: {skipped_low_sources}, "
                        f"sin sugerencia: {skipped_no_suggestion}, "
                        f"score<{min_score:.1f}: {skipped_low_score}, "
                        f"sugerencia sin insumo exacto: {skipped_unresolved}, "
                        f"nombre inválido: {skipped_invalid}."
                    ),
                )
            log_event(
                request.user,
                "AUTO_APPLY_SUGGESTIONS",
                "inventario.InsumoAlias",
                "",
                {
                    "filters": {
                        "cross_q_norm": cross_q_norm_post,
                        "cross_only_suggested": cross_only_suggested_post,
                        "cross_min_sources": cross_min_sources_post,
                        "cross_score_min": float(cross_score_min_post),
                        "cross_point_tipo": cross_point_tipo_post,
                        "cross_source": cross_source_post,
                        "cross_sort_by": cross_sort_by_post,
                        "cross_sort_dir": cross_sort_dir_post,
                        "cross_limit": cross_limit_post,
                        "cross_offset": cross_offset_post,
                    },
                    "thresholds": {
                        "min_score": float(min_score),
                        "min_sources": min_sources,
                        "max_rows": max_rows,
                    },
                    "summary": {
                        "processed": processed,
                        "created": created,
                        "updated": updated,
                        "point_resolved": point_resolved_total,
                        "recetas_resolved": recetas_resolved_total,
                        "skipped_low_sources": skipped_low_sources,
                        "skipped_no_suggestion": skipped_no_suggestion,
                        "skipped_low_score": skipped_low_score,
                        "skipped_unresolved": skipped_unresolved,
                        "skipped_invalid": skipped_invalid,
                    },
                },
            )
        elif action == "reprocess_recetas_pending":
            result = _reprocess_recetas_pending_matching()
            messages.success(
                request,
                (
                    "Reproceso recetas completado. "
                    f"Procesadas: {result['total']}, "
                    f"Auto/OK: {result['auto_ok']}, "
                    f"Por revisar: {result['review']}, "
                    f"Pendientes de integración: {result['rejected']}, "
                    f"Ligadas a insumo: {result['linked']}."
                ),
            )
        elif action == "clear_import_preview":
            request.session.pop("inventario_alias_import_preview", None)
            request.session.pop("inventario_alias_import_stats", None)
            messages.success(request, "Pendientes de importación limpiados.")
        elif action == "master_normalize":
            if not can_view_maestros(request.user):
                raise PermissionDenied("No tienes permisos para normalización de datos maestros.")
            scope, master_q, master_limit, master_offset = _read_master_normalize_filters(request.POST)
            mode = str(request.POST.get("master_mode") or "preview").strip().lower()
            apply_changes = mode == "apply"
            if apply_changes and (not has_any_role(request.user, ROLE_ADMIN, ROLE_DG)):
                messages.error(request, "Solo ADMIN o DG pueden aplicar normalización persistente.")
            else:
                result = _run_master_normalize(
                    scope=scope,
                    q=master_q,
                    limit=master_limit,
                    offset=master_offset,
                    apply_changes=apply_changes,
                )
                if apply_changes:
                    messages.success(
                        request,
                        (
                            "Normalización aplicada. "
                            f"Candidatos: {result['totales']['candidates']}. "
                            f"Cambios detectados: {result['totales']['changed']}. "
                            f"Actualizados: {result['totales']['updated']}."
                        ),
                    )
                else:
                    messages.info(
                        request,
                        (
                            "Vista previa de normalización generada. "
                            f"Candidatos: {result['totales']['candidates']}. "
                            f"Cambios potenciales: {result['totales']['changed']}."
                        ),
                    )
                log_event(
                    request.user,
                    "MASTER_NORMALIZE_APPLY" if apply_changes else "MASTER_NORMALIZE_PREVIEW",
                    "master.Normalization",
                    "",
                    payload={
                        "scope": scope,
                        "q": master_q,
                        "limit": master_limit,
                        "offset": master_offset,
                        "totals": result["totales"],
                    },
                )
        elif action == "resolve_duplicate_insumo":
            source_id_raw = (request.POST.get("source_insumo_id") or "").strip()
            target_id_raw = (request.POST.get("target_insumo_id") or "").strip()
            source = Insumo.objects.filter(pk=source_id_raw).first() if source_id_raw.isdigit() else None
            target = Insumo.objects.filter(pk=target_id_raw).first() if target_id_raw.isdigit() else None
            if not source or not target:
                messages.error(request, "Selecciona origen y destino válidos para resolver duplicado.")
            elif source.id == target.id:
                messages.error(request, "Origen y destino no pueden ser el mismo insumo.")
            else:
                with transaction.atomic():
                    merge_stats = _merge_insumo_into_target(source, target)
                    ok_alias, alias_norm, _ = _upsert_alias(source.nombre, target)
                    if ok_alias:
                        _remove_pending_name_from_session(request, alias_norm)
                        _remove_pending_name_from_recent_runs(alias_norm)
                        _resolve_cross_source_with_alias(source.nombre, target)
                messages.success(
                    request,
                    (
                        f"Duplicado resuelto: '{source.nombre}' → '{target.nombre}'. "
                        f"Referencias {merge_stats['aliases_updated']}, Costos {merge_stats['costos_updated']}, "
                        f"Líneas receta {merge_stats['lineas_updated']}, Movimientos {merge_stats['movimientos_updated']}, "
                        f"Ajustes {merge_stats['ajustes_updated']}, Solicitudes {merge_stats['solicitudes_updated']}."
                    ),
                )
                log_event(
                    request.user,
                    "MASTER_DUPLICATE_RESOLVE_INSUMO",
                    "maestros.Insumo",
                    target.id,
                    payload={
                        "source_id": source.id,
                        "target_id": target.id,
                        "source_nombre": source.nombre,
                        "target_nombre": target.nombre,
                        **merge_stats,
                    },
                )
        elif action == "resolve_duplicate_insumo_group":
            duplicate_key = normalizar_nombre((request.POST.get("duplicate_key") or "").strip())
            target_id_raw = (request.POST.get("target_insumo_id") or "").strip()
            if not duplicate_key:
                messages.error(request, "Llave de duplicado inválida.")
            else:
                members = list(
                    Insumo.objects.filter(nombre_normalizado=duplicate_key).order_by("-activo", "id")
                )
                if len(members) < 2:
                    messages.warning(request, "Ese grupo ya no tiene duplicados activos para resolver.")
                else:
                    target = None
                    if target_id_raw.isdigit():
                        target = next((m for m in members if m.id == int(target_id_raw)), None)
                    if not target:
                        target = members[0]
                    sources = [m for m in members if m.id != target.id]
                    totals = {
                        "sources_resolved": 0,
                        "aliases_updated": 0,
                        "costos_updated": 0,
                        "lineas_updated": 0,
                        "movimientos_updated": 0,
                        "ajustes_updated": 0,
                        "solicitudes_updated": 0,
                        "existencia_merged": 0,
                    }
                    with transaction.atomic():
                        for source in sources:
                            merge_stats = _merge_insumo_into_target(source, target)
                            totals["sources_resolved"] += 1
                            for k, v in merge_stats.items():
                                totals[k] += int(v or 0)
                            ok_alias, alias_norm, _ = _upsert_alias(source.nombre, target)
                            if ok_alias:
                                _remove_pending_name_from_session(request, alias_norm)
                                _remove_pending_name_from_recent_runs(alias_norm)
                                _resolve_cross_source_with_alias(source.nombre, target)

                    messages.success(
                        request,
                        (
                            f"Grupo '{duplicate_key}' resuelto hacia '{target.nombre}'. "
                            f"Fuentes consolidadas {totals['sources_resolved']}. "
                            f"Referencias {totals['aliases_updated']}, Costos {totals['costos_updated']}, "
                            f"Líneas receta {totals['lineas_updated']}."
                        ),
                    )
                    log_event(
                        request.user,
                        "MASTER_DUPLICATE_RESOLVE_GROUP",
                        "maestros.Insumo",
                        target.id,
                        payload={
                            "duplicate_key": duplicate_key,
                            "target_id": target.id,
                            "target_nombre": target.nombre,
                            **totals,
                        },
                    )

        base_url = reverse("inventario:aliases_catalog")
        redirect_params = {}
        if next_q:
            redirect_params["q"] = next_q
        cross_q_post, _, cross_only_suggested_post, cross_min_sources_post, cross_score_min_post = _read_cross_filters(
            request.POST
        )
        cross_point_tipo_post, _ = _read_cross_point_tipo(request.POST)
        cross_source_post = _read_cross_source(request.POST)
        cross_limit_post, cross_offset_post, cross_sort_by_post, cross_sort_dir_post = _read_cross_table_controls(
            request.POST
        )
        if cross_q_post:
            redirect_params["cross_q"] = cross_q_post
        if "cross_min_sources" in request.POST:
            redirect_params["cross_min_sources"] = cross_min_sources_post
        if "cross_score_min" in request.POST:
            redirect_params["cross_score_min"] = cross_score_min_post
        if "cross_limit" in request.POST:
            redirect_params["cross_limit"] = cross_limit_post
        if "cross_offset" in request.POST:
            redirect_params["cross_offset"] = cross_offset_post
        if "cross_sort_by" in request.POST:
            redirect_params["cross_sort_by"] = cross_sort_by_post
        if "cross_sort_dir" in request.POST:
            redirect_params["cross_sort_dir"] = cross_sort_dir_post
        if "cross_point_tipo" in request.POST:
            redirect_params["cross_point_tipo"] = cross_point_tipo_post
        if "cross_source" in request.POST:
            redirect_params["cross_source"] = cross_source_post
        if cross_only_suggested_post:
            redirect_params["cross_only_suggested"] = "1"
        master_scope_post, master_q_post, master_limit_post, master_offset_post = _read_master_normalize_filters(
            request.POST
        )
        if "master_scope" in request.POST:
            redirect_params["master_scope"] = master_scope_post
        if master_q_post:
            redirect_params["master_q"] = master_q_post
        if "master_limit" in request.POST:
            redirect_params["master_limit"] = master_limit_post
        if "master_offset" in request.POST:
            redirect_params["master_offset"] = master_offset_post

        (
            master_dup_scope_post,
            master_dup_q_post,
            master_dup_include_inactive_post,
            master_dup_min_count_post,
            master_dup_limit_post,
            master_dup_offset_post,
        ) = _read_master_duplicates_filters(request.POST)
        if "master_dup_scope" in request.POST:
            redirect_params["master_dup_scope"] = master_dup_scope_post
        if master_dup_q_post:
            redirect_params["master_dup_q"] = master_dup_q_post
        if master_dup_include_inactive_post:
            redirect_params["master_dup_include_inactive"] = "1"
        if "master_dup_min_count" in request.POST:
            redirect_params["master_dup_min_count"] = master_dup_min_count_post
        if "master_dup_limit" in request.POST:
            redirect_params["master_dup_limit"] = master_dup_limit_post
        if "master_dup_offset" in request.POST:
            redirect_params["master_dup_offset"] = master_dup_offset_post
        if redirect_params:
            return redirect(f"{base_url}?{urlencode(redirect_params)}")
        return redirect(base_url)

    q = (request.GET.get("q") or "").strip()
    aliases_qs = InsumoAlias.objects.select_related("insumo").order_by("nombre")
    if q:
        q_norm = normalizar_nombre(q)
        aliases_qs = aliases_qs.filter(
            Q(nombre__icontains=q)
            | Q(nombre_normalizado__icontains=q_norm)
            | Q(insumo__nombre__icontains=q)
        )
    paginator = Paginator(aliases_qs, 100)
    page = paginator.get_page(request.GET.get("page"))

    session_pending = list(request.session.get("inventario_pending_preview", []))[:120]
    hidden_run_id = request.session.get("inventario_hidden_pending_run_id")

    latest_runs = list(
        AlmacenSyncRun.objects.only(
            "id",
            "started_at",
            "source",
            "status",
            "matched",
            "unmatched",
            "rows_stock_read",
            "rows_mov_read",
            "aliases_created",
            "insumos_created",
            "pending_preview",
        ).order_by("-started_at")[:20]
    )
    latest_sync = latest_runs[0] if latest_runs else None
    latest_pending_run = None
    hidden_pending_run = None
    for run in latest_runs:
        if hidden_run_id and run.id == hidden_run_id:
            hidden_pending_run = run
            continue
        if isinstance(run.pending_preview, list) and run.pending_preview:
            latest_pending_run = run
            break

    persisted_pending = list((latest_pending_run.pending_preview if latest_pending_run else [])[:120])
    pending_preview = session_pending or persisted_pending

    recent_runs = latest_runs[:10]
    total_matched = sum(int(r.matched or 0) for r in recent_runs)
    total_unmatched = sum(int(r.unmatched or 0) for r in recent_runs)
    total_rows = total_matched + total_unmatched
    match_rate = round((total_matched * 100.0 / total_rows), 2) if total_rows else 100.0
    ok_runs = sum(1 for r in recent_runs if r.status == AlmacenSyncRun.STATUS_OK)
    pending_recent_runs = [
        {
            "id": r.id,
            "started_at": r.started_at,
            "source_label": r.get_source_display(),
            "status": r.status,
            "matched": int(r.matched or 0),
            "unmatched": int(r.unmatched or 0),
            "has_preview": bool(isinstance(r.pending_preview, list) and r.pending_preview),
            "is_hidden": bool(hidden_run_id and r.id == hidden_run_id),
        }
        for r in recent_runs
    ]

    pending_grouped = _build_pending_grouped(pending_preview)
    auto_default_score = 90.0
    auto_default_min_sources = 2
    cross_point_tipo, point_tipos_filter = _read_cross_point_tipo(request.GET)
    unified_rows, point_unmatched_count, receta_pending_lines = _build_cross_unified_rows(
        pending_grouped,
        point_tipos=point_tipos_filter,
    )
    overlaps = sum(1 for row in unified_rows if int(row.get("sources_active") or 0) >= 2)
    insumo_norm_set = set(
        Insumo.objects.filter(activo=True).values_list("nombre_normalizado", flat=True)
    )
    auto_apply_candidates = sum(
        1
        for row in unified_rows
        if int(row.get("sources_active") or 0) >= auto_default_min_sources
        and row.get("suggestion")
        and float(row.get("score_max") or 0.0) >= auto_default_score
        and normalizar_nombre(str(row.get("suggestion") or "")) in insumo_norm_set
    )

    cross_q, cross_q_norm, cross_only_suggested, cross_min_sources, cross_score_min = _read_cross_filters(request.GET)
    cross_source = _read_cross_source(request.GET)
    cross_filtered_rows = _apply_cross_filters(
        unified_rows,
        cross_q_norm=cross_q_norm,
        cross_only_suggested=cross_only_suggested,
        cross_min_sources=cross_min_sources,
        cross_score_min=cross_score_min,
    )
    if cross_source == "ALMACEN":
        cross_filtered_rows = [row for row in cross_filtered_rows if int(row.get("almacen_count") or 0) > 0]
    elif cross_source == "POINT":
        cross_filtered_rows = [row for row in cross_filtered_rows if int(row.get("point_count") or 0) > 0]
    elif cross_source == "RECETAS":
        cross_filtered_rows = [row for row in cross_filtered_rows if int(row.get("receta_count") or 0) > 0]
    cross_limit, cross_offset, cross_sort_by, cross_sort_dir = _read_cross_table_controls(request.GET)
    cross_filtered_sorted_rows = _sort_cross_rows(
        cross_filtered_rows,
        sort_by=cross_sort_by,
        sort_dir=cross_sort_dir,
    )
    cross_source_stats = {
        "point_rows": sum(1 for row in cross_filtered_sorted_rows if int(row.get("point_count") or 0) > 0),
        "almacen_rows": sum(1 for row in cross_filtered_sorted_rows if int(row.get("almacen_count") or 0) > 0),
        "receta_rows": sum(1 for row in cross_filtered_sorted_rows if int(row.get("receta_count") or 0) > 0),
        "multi_source_rows": sum(1 for row in cross_filtered_sorted_rows if int(row.get("sources_active") or 0) >= 2),
    }

    (
        master_scope,
        master_q,
        master_limit,
        master_offset,
    ) = _read_master_normalize_filters(request.GET)
    master_normalize_preview = _run_master_normalize(
        scope=master_scope,
        q=master_q,
        limit=master_limit,
        offset=master_offset,
        apply_changes=False,
    )

    (
        master_dup_scope,
        master_dup_q,
        master_dup_include_inactive,
        master_dup_min_count,
        master_dup_limit,
        master_dup_offset,
    ) = _read_master_duplicates_filters(request.GET)
    master_duplicates = _build_master_duplicates(
        scope=master_dup_scope,
        q=master_dup_q,
        include_inactive=master_dup_include_inactive,
        min_count=master_dup_min_count,
        limit=master_dup_limit,
        offset=master_dup_offset,
    )
    master_duplicate_insumo_groups = [
        g for g in (master_duplicates.get("items") or []) if g.get("group_type") == "insumos"
    ]

    export_format = (request.GET.get("export") or "").strip().lower()
    if export_format == "cross_pending_csv":
        return _export_cross_pending_csv(cross_filtered_sorted_rows)
    if export_format == "cross_pending_xlsx":
        return _export_cross_pending_xlsx(cross_filtered_sorted_rows)
    if export_format in {"alias_template_csv", "alias_template_xlsx"}:
        return _export_alias_template(export_format)
    if export_format in {"alias_import_preview_csv", "alias_import_preview_xlsx"}:
        preview_rows = list(request.session.get("inventario_alias_import_preview", []))
        if not preview_rows:
            messages.warning(request, "No hay pendientes de importación para exportar.")
            return redirect("inventario:aliases_catalog")
        if export_format == "alias_import_preview_csv":
            return _export_alias_import_preview_csv(preview_rows)
        return _export_alias_import_preview_xlsx(preview_rows)
    if export_format == "aliases_csv":
        return _export_aliases_catalog_csv(aliases_qs)
    if export_format == "aliases_xlsx":
        return _export_aliases_catalog_xlsx(aliases_qs)
    if export_format == "master_duplicates_csv":
        return _export_master_duplicates_csv(master_duplicates["items"])
    if export_format == "master_duplicates_xlsx":
        return _export_master_duplicates_xlsx(master_duplicates["items"])

    import_preview = list(request.session.get("inventario_alias_import_preview", []))[:200]
    import_stats = request.session.get("inventario_alias_import_stats", {})
    cross_rows_total = len(cross_filtered_sorted_rows)
    cross_unified_rows = cross_filtered_sorted_rows[cross_offset : cross_offset + cross_limit]
    cross_has_prev = cross_offset > 0
    cross_has_next = (cross_offset + len(cross_unified_rows)) < cross_rows_total
    cross_prev_offset = max(cross_offset - cross_limit, 0)
    cross_next_offset = cross_offset + cross_limit if cross_has_next else cross_offset
    cross_query_common = urlencode(
        {
            "cross_q": cross_q,
            "cross_min_sources": cross_min_sources,
            "cross_score_min": cross_score_min,
            "cross_point_tipo": cross_point_tipo,
            "cross_source": cross_source,
            "cross_sort_by": cross_sort_by,
            "cross_sort_dir": cross_sort_dir,
            "cross_limit": cross_limit,
            **({"cross_only_suggested": "1"} if cross_only_suggested else {}),
        }
    )
    cross_query_source_todos = urlencode(
        {
            "cross_q": cross_q,
            "cross_source": "TODOS",
            "cross_min_sources": cross_min_sources,
            "cross_score_min": cross_score_min,
            "cross_point_tipo": cross_point_tipo,
            "cross_sort_by": cross_sort_by,
            "cross_sort_dir": cross_sort_dir,
            "cross_limit": cross_limit,
            "cross_offset": 0,
            **({"cross_only_suggested": "1"} if cross_only_suggested else {}),
        }
    )
    cross_query_source_almacen = urlencode(
        {
            "cross_q": cross_q,
            "cross_source": "ALMACEN",
            "cross_min_sources": cross_min_sources,
            "cross_score_min": cross_score_min,
            "cross_point_tipo": cross_point_tipo,
            "cross_sort_by": cross_sort_by,
            "cross_sort_dir": cross_sort_dir,
            "cross_limit": cross_limit,
            "cross_offset": 0,
            **({"cross_only_suggested": "1"} if cross_only_suggested else {}),
        }
    )
    cross_query_source_point = urlencode(
        {
            "cross_q": cross_q,
            "cross_source": "POINT",
            "cross_min_sources": cross_min_sources,
            "cross_score_min": cross_score_min,
            "cross_point_tipo": cross_point_tipo,
            "cross_sort_by": cross_sort_by,
            "cross_sort_dir": cross_sort_dir,
            "cross_limit": cross_limit,
            "cross_offset": 0,
            **({"cross_only_suggested": "1"} if cross_only_suggested else {}),
        }
    )
    cross_query_source_recetas = urlencode(
        {
            "cross_q": cross_q,
            "cross_source": "RECETAS",
            "cross_min_sources": cross_min_sources,
            "cross_score_min": cross_score_min,
            "cross_point_tipo": cross_point_tipo,
            "cross_sort_by": cross_sort_by,
            "cross_sort_dir": cross_sort_dir,
            "cross_limit": cross_limit,
            "cross_offset": 0,
            **({"cross_only_suggested": "1"} if cross_only_suggested else {}),
        }
    )
    pending_source_buckets: dict[str, dict[str, object]] = {}
    for row in pending_grouped:
        sources = [item.strip() for item in str(row.get("sources") or "").split(",") if item.strip() and item.strip() != "-"]
        if not sources:
            sources = ["sin fuente"]
        for source in sources:
            bucket = pending_source_buckets.setdefault(
                source,
                {"source": source, "count": 0, "top_name": row.get("nombre_origen") or "", "top_score": 0.0},
            )
            bucket["count"] += int(row.get("count") or 0)
            score_max = float(row.get("score_max") or 0.0)
            if score_max >= float(bucket["top_score"] or 0.0):
                bucket["top_name"] = row.get("nombre_origen") or bucket["top_name"]
                bucket["top_score"] = score_max
    pending_source_cards = sorted(
        pending_source_buckets.values(),
        key=lambda item: (-int(item["count"]), str(item["source"])),
    )
    pending_focus_rows = pending_grouped[:3]
    if pending_focus_rows:
        top_pending = pending_focus_rows[0]
        top_pending_sources = [item.strip() for item in str(top_pending.get("sources") or "").split(",") if item.strip() and item.strip() != "-"]
        pending_focus = {
            "label": f"{top_pending.get('count', 0)} filas · {', '.join(top_pending_sources) if top_pending_sources else 'sin fuente'}",
            "summary": (
                f"La integración sigue condicionada por '{top_pending.get('nombre_origen')}' "
                "como nombre fuera de maestro entre fuentes operativas."
            ),
            "action_label": "Abrir referencias filtradas",
            "action_detail": "Consolida el nombre origen antes de reprocesar inventario, recetas o pendientes del sistema comercial.",
            "action_url": f"{reverse('inventario:aliases_catalog')}?q={quote_plus(str(top_pending.get('nombre_origen') or ''))}",
            "tone": "warning",
        }
    else:
        pending_focus = {
            "label": "Integración al día",
            "summary": "No hay nombres pendientes de integración en el backlog operativo actual.",
            "action_label": "Revisar referencias",
            "action_detail": "Puedes revisar referencias de forma preventiva para mantener el catálogo limpio.",
            "action_url": reverse("inventario:aliases_catalog"),
            "tone": "success",
        }

    references_workflow_rows = [
        {
            "step": "01",
            "title": "Entrada operativa",
            "detail": "Almacén, recetas y fuente comercial externa siguen alimentando nombres y códigos al ERP.",
            "open": len(pending_preview),
            "closed": max(total_matched, 0),
            "tone": "warning" if len(pending_preview) else "success",
            "owner": "Inventario / Sucursales",
            "completion": 100 if len(pending_preview) == 0 else 35,
            "next_step": "Depurar backlog de nombres origen" if len(pending_preview) else "Entrada estabilizada",
            "action_label": "Abrir backlog",
            "action_href": reverse("inventario:aliases_catalog"),
        },
        {
            "step": "02",
            "title": "Resolución maestra",
            "detail": "Cada nombre origen debe consolidarse contra un solo artículo maestro del ERP.",
            "open": total_unmatched,
            "closed": max(total_matched, 0),
            "tone": "warning" if total_unmatched else "success",
            "owner": "Maestros / Inventario",
            "completion": 100 if total_unmatched == 0 else 70,
            "next_step": "Cerrar referencias maestras pendientes" if total_unmatched else "Catálogo maestro estabilizado",
            "action_label": "Resolver referencias",
            "action_href": reverse("inventario:aliases_catalog"),
        },
        {
            "step": "03",
            "title": "Liberación operativa",
            "detail": "Solo artículos maestros liberados pueden pasar a inventario, costeo, compras y reabasto.",
            "open": receta_pending_lines + point_unmatched_count,
            "closed": overlaps,
            "tone": "warning" if (receta_pending_lines + point_unmatched_count) else "success",
            "owner": "ERP / Operaciones",
            "completion": 100 if (receta_pending_lines + point_unmatched_count) == 0 else 85,
            "next_step": "Liberar impacto operativo por módulo" if (receta_pending_lines + point_unmatched_count) else "Circuito operativo cerrado",
            "action_label": "Ver impacto ERP",
            "action_href": reverse("dashboard"),
        },
    ]
    references_release_summary = {
        "label": "Catálogo liberado" if total_unmatched == 0 else "Catálogo con brechas abiertas",
        "tone": "success" if total_unmatched == 0 else "warning",
        "detail": (
            "No hay referencias maestras pendientes; inventario puede operar íntegro con artículo maestro."
            if total_unmatched == 0
            else f"Quedan {total_unmatched} referencia(s) maestras pendientes antes del cierre operativo del catálogo."
        ),
    }
    reference_module_cards = [
        {
            "title": "Inventario",
            "tone": "warning" if len(pending_preview) else "success",
            "detail": (
                f"{len(pending_preview)} referencia(s) abiertas siguen afectando existencias, movimientos o merma."
                if pending_preview
                else "Inventario ya opera con referencias cerradas en la carga vigente."
            ),
            "action_label": "Abrir inventario",
            "action_href": reverse("inventario:existencias"),
        },
        {
            "title": "Recetas / BOM",
            "tone": "warning" if receta_pending_lines else "success",
            "detail": (
                f"{receta_pending_lines} componente(s) de receta siguen abiertos frente al artículo ERP."
                if receta_pending_lines
                else "Recetas ya consumen artículos ERP sin referencias abiertas."
            ),
            "action_label": "Abrir recetas",
            "action_href": reverse("recetas:recetas_list"),
        },
        {
            "title": "Fuente comercial",
            "tone": "warning" if point_unmatched_count else "success",
            "detail": (
                f"{point_unmatched_count} referencia(s) comerciales siguen abiertas y afectan homologación cruzada."
                if point_unmatched_count
                else "La fuente comercial ya quedó alineada con el artículo ERP."
            ),
            "action_label": "Abrir centro comercial",
            "action_href": reverse("maestros:point_pending_review"),
        },
        {
            "title": "Compras y reabasto",
            "tone": "warning" if total_unmatched else "success",
            "detail": (
                f"{total_unmatched} referencia(s) maestras abiertas aún pueden distorsionar compras, reabasto y análisis."
                if total_unmatched
                else "Compras y reabasto ya operan con referencias ERP estables."
            ),
            "action_label": "Abrir compras",
            "action_href": reverse("compras:solicitudes"),
        },
    ]
    references_next_step_summary = (
        {
            "title": "Siguiente paso recomendado",
            "label": "Catálogo ERP liberado",
            "detail": "No hay brechas abiertas; puedes operar inventario, compras y recetas sin referencias pendientes.",
            "action_label": "Abrir existencias",
            "action_href": reverse("inventario:existencias"),
            "tone": "success",
        }
        if total_unmatched == 0
        else {
            "title": "Siguiente paso recomendado",
            "label": "Cerrar referencias maestras abiertas",
            "detail": (
                f"Empieza por '{top_pending.get('nombre_origen')}' para reducir el bloqueo dominante del catálogo."
                if pending_focus_rows
                else "Revisa el catálogo y consolida las referencias abiertas antes del siguiente reproceso."
            ),
            "action_label": "Abrir catálogo ERP",
            "action_href": reverse("inventario:aliases_catalog"),
            "tone": "warning",
        }
    )
    erp_governance_rows = [
        {
            "front": "Fuente operativa",
            "owner": "Inventario / Sucursales",
            "blockers": len(pending_preview),
            "completion": 100 if len(pending_preview) == 0 else 35,
            "detail": (
                "Las fuentes operativas ya llegan sin nombres abiertos."
                if len(pending_preview) == 0
                else f"{len(pending_preview)} referencia(s) siguen entrando con nombre abierto."
            ),
            "next_step": (
                "Monitoreo preventivo."
                if len(pending_preview) == 0
                else "Cerrar backlog de nombres origen."
            ),
            "url": reverse("inventario:aliases_catalog"),
            "cta": "Abrir backlog",
        },
        {
            "front": "Artículo maestro",
            "owner": "Maestros / Inventario",
            "blockers": total_unmatched,
            "completion": 100 if total_unmatched == 0 else 70,
            "detail": (
                "El artículo maestro ya quedó estable para operación."
                if total_unmatched == 0
                else f"{total_unmatched} referencia(s) todavía requieren consolidación maestra."
            ),
            "next_step": (
                "Mantener consolidación preventiva."
                if total_unmatched == 0
                else "Resolver referencias maestras pendientes."
            ),
            "url": reverse("inventario:aliases_catalog"),
            "cta": "Resolver referencias",
        },
        {
            "front": "Impacto ERP",
            "owner": "ERP / Operaciones",
            "blockers": receta_pending_lines + point_unmatched_count,
            "completion": 100 if (receta_pending_lines + point_unmatched_count) == 0 else 85,
            "detail": (
                "Recetas, inventario y compras ya operan con referencias cerradas."
                if (receta_pending_lines + point_unmatched_count) == 0
                else f"{receta_pending_lines + point_unmatched_count} impacto(s) siguen abiertos en operación."
            ),
            "next_step": (
                "Mantener circuito estable."
                if (receta_pending_lines + point_unmatched_count) == 0
                else "Liberar el impacto operativo por módulo."
            ),
            "url": reverse("dashboard"),
            "cta": "Ver impacto ERP",
        },
    ]

    context = {
        "q": q,
        "page": page,
        "pending_preview": pending_preview,
        "pending_grouped": pending_grouped[:80],
        "pending_source": "session" if session_pending else ("persisted" if persisted_pending else ""),
        "latest_pending_run": latest_pending_run,
        "hidden_pending_run": hidden_pending_run,
        "hidden_run_id": hidden_run_id,
        "latest_sync": latest_sync,
        "matching_summary": {
            "runs_count": len(recent_runs),
            "ok_runs": ok_runs,
            "total_matched": total_matched,
            "total_unmatched": total_unmatched,
            "match_rate": match_rate,
        },
        "pending_recent_runs": pending_recent_runs,
        "pending_visible_count": len(pending_preview),
        "pending_unique_count": len(pending_grouped),
        "pending_source_cards": pending_source_cards,
        "pending_focus_rows": pending_focus_rows,
        "pending_focus": pending_focus,
        "auto_default_score": auto_default_score,
        "auto_default_min_sources": auto_default_min_sources,
        "auto_default_limit": 80,
        "auto_apply_candidates": auto_apply_candidates,
        "insumo_alias_targets": canonicalized_insumo_selector(limit=1200),
        "can_manage_inventario": can_manage_inventario(request.user),
        "cross_q": cross_q,
        "cross_min_sources": cross_min_sources,
        "cross_score_min": cross_score_min,
        "cross_point_tipo": cross_point_tipo,
        "cross_source": cross_source,
        "cross_source_stats": cross_source_stats,
        "cross_only_suggested": cross_only_suggested,
        "cross_filtered_count": len(cross_filtered_sorted_rows),
        "cross_total_count": len(unified_rows),
        "cross_limit": cross_limit,
        "cross_offset": cross_offset,
        "cross_sort_by": cross_sort_by,
        "cross_sort_dir": cross_sort_dir,
        "cross_returned_count": len(cross_unified_rows),
        "cross_has_prev": cross_has_prev,
        "cross_has_next": cross_has_next,
        "cross_prev_offset": cross_prev_offset,
        "cross_next_offset": cross_next_offset,
        "cross_query_common": cross_query_common,
        "cross_query_source_todos": cross_query_source_todos,
        "cross_query_source_almacen": cross_query_source_almacen,
        "cross_query_source_point": cross_query_source_point,
        "cross_query_source_recetas": cross_query_source_recetas,
        "cross_summary": {
            "point_unmatched": point_unmatched_count,
            "almacen_unmatched": len(pending_preview),
            "recetas_unmatched": receta_pending_lines,
            "overlaps": overlaps,
        },
        "cross_unified_rows": cross_unified_rows,
        "alias_import_preview": import_preview,
        "alias_import_stats": import_stats,
        "master_normalize": master_normalize_preview,
        "master_scope": master_scope,
        "master_q": master_q,
        "master_limit": master_limit,
        "master_offset": master_offset,
        "master_duplicates": master_duplicates,
        "master_dup_scope": master_dup_scope,
        "master_dup_q": master_dup_q,
        "master_dup_include_inactive": master_dup_include_inactive,
        "master_dup_min_count": master_dup_min_count,
        "master_dup_limit": master_dup_limit,
        "master_dup_offset": master_dup_offset,
        "master_duplicate_insumo_groups": master_duplicate_insumo_groups,
        "references_workflow_rows": references_workflow_rows,
        "references_release_summary": references_release_summary,
        "reference_module_cards": reference_module_cards,
        "references_next_step_summary": references_next_step_summary,
        "erp_governance_rows": erp_governance_rows,
        "executive_radar_rows": _inventario_executive_radar_rows(erp_governance_rows),
    }
    context["erp_command_center"] = _inventario_command_center(
        governance_rows=erp_governance_rows,
        maturity_summary={
            "attention_steps": sum(1 for row in references_workflow_rows if row.get("tone") != "success"),
            "next_priority_title": references_next_step_summary.get("label", "Catálogo ERP estable"),
            "next_priority_detail": references_next_step_summary.get("detail", ""),
            "next_priority_url": references_next_step_summary.get("action_href", reverse("inventario:aliases_catalog")),
            "next_priority_cta": references_next_step_summary.get("action_label", "Abrir catálogo ERP"),
        },
    )
    return render(request, "inventario/aliases_catalog.html", context)


@login_required
def sync_drive_now(request: HttpRequest) -> HttpResponse:
    if request.method != "POST":
        return redirect("dashboard")
    if not can_manage_inventario(request.user):
        raise PermissionDenied("No tienes permisos para ejecutar sincronización manual.")

    next_url = (request.POST.get("next") or "").strip()
    if not next_url.startswith("/"):
        next_url = ""

    try:
        cmd = [
            sys.executable,
            "manage.py",
            "sync_almacen_drive",
            "--create-missing-insumos",
        ]
        subprocess.Popen(
            cmd,
            cwd=str(settings.BASE_DIR),
            env=os.environ.copy(),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.STDOUT,
            start_new_session=True,
        )
        messages.success(
            request,
            "Sincronización manual iniciada. Refresca el dashboard en 1-2 minutos para ver el resultado.",
        )
    except Exception as exc:
        messages.error(request, f"No se pudo iniciar la sincronización manual: {exc}")

    return redirect(next_url or "dashboard")


def _inventario_enterprise_chain(
    *,
    focus: str,
    total_rows: int,
    blocked_count: int,
    critical_count: int = 0,
    pending_count: int = 0,
) -> list[dict[str, object]]:
    stage_meta = {
        "Maestro ERP": {
            "owner": "Maestros / Inventario",
            "next_step": "Cerrar unidad, categoría, proveedor y artículo canónico.",
        },
        "Existencia operativa": {
            "owner": "Inventario / Almacén",
            "next_step": "Validar stock, mínimos y consumo diario antes del reabasto.",
        },
        "Reabasto": {
            "owner": "Compras / Producción",
            "next_step": "Convertir alertas críticas en compra o producción formal.",
        },
        "Bitácora de movimientos": {
            "owner": "Inventario / Operación",
            "next_step": "Mantener ledger único y trazable por artículo canónico.",
        },
        "Impacto en existencia": {
            "owner": "Inventario / Planeación",
            "next_step": "Corregir referencias o bloqueos antes de afectar stock.",
        },
        "Ajustes pendientes": {
            "owner": "Inventario / Auditoría",
            "next_step": "Aprobar o rechazar ajustes antes del cierre del día.",
        },
        "Existencia reconciliada": {
            "owner": "Inventario / DG",
            "next_step": "Mantener historial aplicado y diferencia conciliada.",
        },
        "Alertas priorizadas": {
            "owner": "Inventario / Compras",
            "next_step": "Priorizar quiebres y artículos bajo reorden.",
        },
        "Acción de reabasto": {
            "owner": "Compras / Reabasto",
            "next_step": "Emitir solicitud o corregir inventario según criticidad.",
        },
    }
    if focus == "existencias":
        chain = [
            {
                "step": "01",
                "title": "Maestro ERP",
                "count": blocked_count,
                "status": "Bloqueos" if blocked_count else "Listo",
                "tone": "warning" if blocked_count else "success",
                "detail": "Unidad, categoría, proveedor y artículo canónico deben quedar cerrados antes del cálculo de stock.",
                "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
                "cta": "Abrir maestro",
            },
            {
                "step": "02",
                "title": "Existencia operativa",
                "count": total_rows,
                "status": "Control activo",
                "tone": "primary",
                "detail": "Stock, mínimo, máximo y consumo diario consolidados por artículo canónico.",
                "url": reverse("inventario:existencias"),
                "cta": "Ver existencias",
            },
            {
                "step": "03",
                "title": "Reabasto",
                "count": critical_count,
                "status": "Por actuar" if critical_count else "Alineado",
                "tone": "danger" if critical_count else "success",
                "detail": "El reorden se libera cuando maestro y existencias quedan consistentes.",
                "url": reverse("inventario:alertas") + "?nivel=alerta",
                "cta": "Revisar alertas",
            },
        ]
    elif focus == "movimientos":
        chain = [
            {
                "step": "01",
                "title": "Maestro ERP",
                "count": blocked_count,
                "status": "Bloqueos" if blocked_count else "Listo",
                "tone": "warning" if blocked_count else "success",
                "detail": "Sin maestro cerrado, el movimiento pierde trazabilidad y uso operativo.",
                "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
                "cta": "Abrir maestro",
            },
            {
                "step": "02",
                "title": "Bitácora de movimientos",
                "count": total_rows,
                "status": "Documentada" if total_rows else "Sin registros",
                "tone": "primary" if total_rows else "warning",
                "detail": "Entradas, salidas y consumos deben quedar en un único ledger por artículo canónico.",
                "url": reverse("inventario:movimientos"),
                "cta": "Ver bitácora",
            },
            {
                "step": "03",
                "title": "Impacto en existencia",
                "count": pending_count,
                "status": "Por revisar" if pending_count else "Sin pendientes",
                "tone": "warning" if pending_count else "success",
                "detail": "Revisa artículos con faltantes de maestro o consolidación pendiente antes de afectar inventario.",
                "url": reverse("inventario:existencias"),
                "cta": "Ver existencia",
            },
        ]
    elif focus == "ajustes":
        chain = [
            {
                "step": "01",
                "title": "Maestro ERP",
                "count": blocked_count,
                "status": "Bloqueos" if blocked_count else "Listo",
                "tone": "warning" if blocked_count else "success",
                "detail": "Un ajuste sólo debe aplicarse cuando el artículo está listo y trazable.",
                "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
                "cta": "Abrir maestro",
            },
            {
                "step": "02",
                "title": "Ajustes pendientes",
                "count": pending_count,
                "status": "Por aprobar" if pending_count else "Sin pendientes",
                "tone": "warning" if pending_count else "success",
                "detail": "Ajustes documentados pendientes de revisión o aplicación.",
                "url": reverse("inventario:ajustes"),
                "cta": "Ver ajustes",
            },
            {
                "step": "03",
                "title": "Existencia reconciliada",
                "count": total_rows - pending_count if total_rows >= pending_count else 0,
                "status": "Con historial",
                "tone": "primary",
                "detail": "El ERP debe distinguir entre ajuste solicitado y ajuste ya aplicado.",
                "url": reverse("inventario:existencias"),
                "cta": "Ver stock",
            },
        ]
    else:
        chain = [
            {
                "step": "01",
                "title": "Maestro ERP",
                "count": blocked_count,
                "status": "Bloqueos" if blocked_count else "Listo",
                "tone": "warning" if blocked_count else "success",
                "detail": "Las alertas son más confiables cuando el maestro está completo.",
                "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
                "cta": "Abrir maestro",
            },
            {
                "step": "02",
                "title": "Alertas priorizadas",
                "count": total_rows,
                "status": "Activas" if total_rows else "Sin alertas",
                "tone": "danger" if critical_count else "warning" if total_rows else "success",
                "detail": "Se priorizan insumos bajo reorden o sin stock con gobierno maestro visible.",
                "url": reverse("inventario:alertas") + "?nivel=all",
                "cta": "Ver alertas",
            },
            {
                "step": "03",
                "title": "Acción de reabasto",
                "count": critical_count,
                "status": "Inmediata" if critical_count else "Controlada",
                "tone": "danger" if critical_count else "success",
                "detail": "Las alertas críticas deben mover reorden, compra o corrección de inventario.",
                "url": reverse("compras:solicitudes"),
                "cta": "Ir a compras",
            },
        ]
    for index, item in enumerate(chain):
        meta = stage_meta.get(str(item.get("title", "")), {})
        item["owner"] = meta.get("owner", "Operación ERP")
        item["next_step"] = meta.get("next_step", "Revisar etapa")
        item["completion"] = 100 if item.get("tone") == "success" else (60 if item.get("tone") in {"warning", "primary"} else 25)
        previous = chain[index - 1] if index else None
        item["depends_on"] = previous["title"] if previous else "Origen del módulo"
        if previous:
            item["dependency_status"] = (
                f"Condicionado por {previous['title'].lower()}"
                if previous.get("tone") != "success"
                else f"Listo desde {previous['title'].lower()}"
            )
        else:
            item["dependency_status"] = "Punto de arranque del módulo"
    return chain


def _inventario_document_stage_rows(
    *,
    focus: str,
    total_rows: int,
    blocked_count: int,
    critical_count: int = 0,
    pending_count: int = 0,
    healthy_count: int = 0,
) -> list[dict[str, object]]:
    if focus == "existencias":
        return [
            {
                "label": "Maestro ERP",
                "owner": "Maestros / DG",
                "open": blocked_count,
                "closed": max(total_rows - blocked_count, 0),
                "completion": round((max(total_rows - blocked_count, 0) / max(total_rows, 1)) * 100) if total_rows else 0,
                "detail": "Artículos listos frente a artículos que siguen bloqueando inventario.",
                "next_step": "Corregir maestro",
                "url": reverse("maestros:insumo_list"),
            },
            {
                "label": "Existencia consolidada",
                "owner": "Inventario / Almacén",
                "open": total_rows,
                "closed": 0,
                "completion": 0,
                "detail": "Filas activas de stock por artículo canónico.",
                "next_step": "Revisar existencias",
                "url": reverse("inventario:existencias"),
            },
            {
                "label": "Reabasto en riesgo",
                "owner": "Compras / Producción",
                "open": critical_count,
                "closed": healthy_count,
                "completion": round((healthy_count / max(critical_count + healthy_count, 1)) * 100) if (critical_count + healthy_count) else 0,
                "detail": "Críticos y bajo reorden frente a artículos ya sanos.",
                "next_step": "Abrir alertas",
                "url": reverse("inventario:alertas") + "?nivel=alerta",
            },
        ]
    if focus == "movimientos":
        return [
            {
                "label": "Maestro ERP",
                "owner": "Maestros / DG",
                "open": blocked_count,
                "closed": max(total_rows - blocked_count, 0),
                "completion": round((max(total_rows - blocked_count, 0) / max(total_rows, 1)) * 100) if total_rows else 0,
                "detail": "Artículos listos para mover contra artículos con bloqueo maestro.",
                "next_step": "Corregir maestro",
                "url": reverse("maestros:insumo_list"),
            },
            {
                "label": "Movimientos documentados",
                "owner": "Inventario / Almacén",
                "open": total_rows,
                "closed": 0,
                "completion": 0,
                "detail": "Eventos registrados en la bitácora operativa.",
                "next_step": "Revisar kardex",
                "url": reverse("inventario:movimientos"),
            },
            {
                "label": "Consolidación maestra",
                "owner": "Maestros / DG",
                "open": pending_count,
                "closed": max(total_rows - pending_count, 0),
                "completion": round((max(total_rows - pending_count, 0) / max(total_rows, 1)) * 100) if total_rows else 0,
                "detail": "Movimientos con artículo canónico frente a movimientos con referencia pendiente.",
                "next_step": "Consolidar referencias",
                "url": reverse("maestros:insumo_list") + "?canonical_status=variantes",
            },
        ]
    if focus == "ajustes":
        return [
            {
                "label": "Maestro ERP",
                "owner": "Maestros / DG",
                "open": blocked_count,
                "closed": max(total_rows - blocked_count, 0),
                "completion": round((max(total_rows - blocked_count, 0) / max(total_rows, 1)) * 100) if total_rows else 0,
                "detail": "Artículos listos para ajuste contra artículos con bloqueo maestro.",
                "next_step": "Corregir maestro",
                "url": reverse("maestros:insumo_list"),
            },
            {
                "label": "Ajustes pendientes",
                "owner": "Inventario / Almacén",
                "open": pending_count,
                "closed": max(total_rows - pending_count, 0),
                "completion": round((max(total_rows - pending_count, 0) / max(total_rows, 1)) * 100) if total_rows else 0,
                "detail": "Ajustes aún no aplicados frente a ajustes ya cerrados.",
                "next_step": "Aplicar ajustes",
                "url": reverse("inventario:ajustes"),
            },
            {
                "label": "Stock reconciliado",
                "owner": "Inventario / Almacén",
                "open": total_rows,
                "closed": max(total_rows - pending_count, 0),
                "completion": round((max(total_rows - pending_count, 0) / max(total_rows, 1)) * 100) if total_rows else 0,
                "detail": "Ajustes con impacto ya trazado sobre existencia.",
                "next_step": "Revisar existencias",
                "url": reverse("inventario:existencias"),
            },
        ]
    return [
        {
            "label": "Maestro ERP",
            "owner": "Maestros / DG",
            "open": blocked_count,
            "closed": max(total_rows - blocked_count, 0),
            "completion": round((max(total_rows - blocked_count, 0) / max(total_rows, 1)) * 100) if total_rows else 0,
            "detail": "Artículos listos frente a artículos que todavía distorsionan alertas.",
            "next_step": "Corregir maestro",
            "url": reverse("maestros:insumo_list"),
        },
        {
            "label": "Alertas abiertas",
            "owner": "Inventario / Almacén",
            "open": total_rows,
            "closed": healthy_count,
            "completion": round((healthy_count / max(total_rows + healthy_count, 1)) * 100) if (total_rows + healthy_count) else 0,
            "detail": "Insumos críticos o bajo reorden frente a insumos sanos.",
            "next_step": "Atender alertas",
            "url": reverse("inventario:alertas") + "?nivel=all",
        },
        {
            "label": "Acción de reabasto",
            "owner": "Compras / Producción",
            "open": critical_count,
            "closed": max(total_rows - critical_count, 0),
            "completion": round((max(total_rows - critical_count, 0) / max(total_rows, 1)) * 100) if total_rows else 0,
            "detail": "Casos que deben moverse a compra, ajuste o reorden.",
            "next_step": "Crear reabasto",
            "url": reverse("compras:solicitudes"),
        },
    ]


def _inventario_operational_health_cards(
    *,
    focus: str,
    total_rows: int,
    blocked_count: int = 0,
    critical_count: int = 0,
    pending_count: int = 0,
    healthy_count: int = 0,
) -> list[dict[str, object]]:
    if focus == "existencias":
        return [
            {
                "label": "Maestro ERP",
                "count": blocked_count,
                "tone": "warning" if blocked_count else "success",
                "detail": (
                    "Hay artículos bloqueando reorden y conciliación."
                    if blocked_count
                    else "El maestro no bloquea la lectura de existencias."
                ),
                "cta": "Corregir maestro" if blocked_count else "Revisar maestro",
                "url": reverse("maestros:insumo_list") + f"?{urlencode({'usage_scope': 'inventory', 'enterprise_status': 'incompletos'})}" if blocked_count else reverse("maestros:insumo_list"),
            },
            {
                "label": "Reabasto",
                "count": critical_count,
                "tone": "danger" if critical_count else "success",
                "detail": (
                    "Hay existencias por debajo de reorden."
                    if critical_count
                    else "Las existencias visibles no muestran urgencia de reabasto."
                ),
                "cta": "Abrir alertas",
                "url": reverse("inventario:alertas"),
            },
            {
                "label": "Referencias",
                "count": pending_count,
                "tone": "warning" if pending_count else "success",
                "detail": (
                    "Hay referencias por consolidar en inventario."
                    if pending_count
                    else "No hay referencias maestras pendientes en esta vista."
                ),
                "cta": "Abrir referencias",
                "url": reverse("inventario:aliases_catalog"),
            },
            {
                "label": "Cobertura sana",
                "count": healthy_count,
                "tone": "success" if healthy_count else "neutral",
                "detail": "Artículos visibles con cobertura suficiente frente al punto de reorden.",
                "cta": "Ver existencias",
                "url": reverse("inventario:existencias"),
            },
        ]
    if focus == "movimientos":
        return [
            {
                "label": "Trazabilidad",
                "count": total_rows,
                "tone": "success" if total_rows else "neutral",
                "detail": "Movimientos visibles con referencia operativa y lectura ERP.",
                "cta": "Ver bitácora",
                "url": reverse("inventario:movimientos"),
            },
            {
                "label": "Maestro ERP",
                "count": blocked_count,
                "tone": "warning" if blocked_count else "success",
                "detail": (
                    "Hay artículos que limitan el registro correcto de movimientos."
                    if blocked_count
                    else "El maestro no bloquea la captura de movimientos."
                ),
                "cta": "Corregir maestro" if blocked_count else "Revisar maestro",
                "url": reverse("maestros:insumo_list") + f"?{urlencode({'usage_scope': 'inventory', 'enterprise_status': 'incompletos'})}" if blocked_count else reverse("maestros:insumo_list"),
            },
            {
                "label": "Referencias",
                "count": pending_count,
                "tone": "warning" if pending_count else "success",
                "detail": (
                    "Hay variantes que conviene consolidar para no dividir el historial."
                    if pending_count
                    else "No hay referencias pendientes dentro de los movimientos visibles."
                ),
                "cta": "Abrir referencias",
                "url": reverse("inventario:aliases_catalog"),
            },
        ]
    if focus == "ajustes":
        return [
            {
                "label": "Pendientes de aprobación",
                "count": pending_count,
                "tone": "warning" if pending_count else "success",
                "detail": (
                    "Hay ajustes que requieren revisión o aplicación."
                    if pending_count
                    else "No hay ajustes pendientes en la vista actual."
                ),
                "cta": "Ver ajustes",
                "url": reverse("inventario:ajustes"),
            },
            {
                "label": "Maestro ERP",
                "count": blocked_count,
                "tone": "warning" if blocked_count else "success",
                "detail": (
                    "Hay artículos incompletos afectando la trazabilidad del ajuste."
                    if blocked_count
                    else "El maestro no bloquea la gestión de ajustes."
                ),
                "cta": "Corregir maestro" if blocked_count else "Revisar maestro",
                "url": reverse("maestros:insumo_list") + f"?{urlencode({'usage_scope': 'inventory', 'enterprise_status': 'incompletos'})}" if blocked_count else reverse("maestros:insumo_list"),
            },
            {
                "label": "Control documental",
                "count": total_rows,
                "tone": "neutral",
                "detail": "Ajustes visibles dentro del circuito de solicitud, revisión y aplicación.",
                "cta": "Abrir ajustes",
                "url": reverse("inventario:ajustes"),
            },
        ]
    return [
        {
            "label": "Críticos",
            "count": critical_count,
            "tone": "danger" if critical_count else "success",
            "detail": (
                "Hay artículos sin stock o en nivel crítico."
                if critical_count
                else "No hay alertas críticas visibles."
            ),
            "cta": "Ver críticos",
            "url": reverse("inventario:alertas") + "?nivel=critico",
        },
        {
            "label": "Reorden",
            "count": pending_count,
            "tone": "warning" if pending_count else "success",
            "detail": (
                "Hay referencias o alertas que siguen abiertas por resolver."
                if pending_count
                else "No hay alertas abiertas de referencias en esta vista."
            ),
            "cta": "Abrir alertas",
            "url": reverse("inventario:alertas"),
        },
        {
            "label": "Maestro ERP",
            "count": blocked_count,
            "tone": "warning" if blocked_count else "success",
            "detail": (
                "Hay artículos incompletos afectando la priorización."
                if blocked_count
                else "El maestro no bloquea la priorización de alertas."
            ),
            "cta": "Corregir maestro" if blocked_count else "Revisar maestro",
            "url": reverse("maestros:insumo_list") + f"?{urlencode({'usage_scope': 'inventory', 'enterprise_status': 'incompletos'})}" if blocked_count else reverse("maestros:insumo_list"),
        },
        {
            "label": "Cobertura sana",
            "count": healthy_count,
            "tone": "success" if healthy_count else "neutral",
            "detail": "Artículos visibles que hoy no exigen acción correctiva.",
            "cta": "Ver suficientes",
            "url": reverse("inventario:alertas") + "?nivel=ok",
        },
    ]


def _inventario_maturity_summary(*, chain: list[dict], default_url: str) -> dict[str, object]:
    completed_steps = sum(1 for item in chain if item.get("tone") == "success")
    attention_steps = len(chain) - completed_steps
    coverage_pct = int(round((completed_steps / len(chain)) * 100)) if chain else 0
    next_priority = next((item for item in chain if item.get("tone") != "success"), None)
    if not next_priority:
        next_priority = {
            "title": "Operación estabilizada",
            "detail": "No hay brechas críticas abiertas en inventario.",
            "url": default_url,
            "cta": "Abrir inventario",
        }
    return {
        "completed_steps": completed_steps,
        "attention_steps": attention_steps,
        "coverage_pct": coverage_pct,
        "next_priority_title": next_priority.get("title", "Operación estabilizada"),
        "next_priority_detail": next_priority.get("detail", "No hay brechas críticas abiertas en inventario."),
        "next_priority_url": next_priority.get("url", default_url),
        "next_priority_cta": next_priority.get("cta", "Abrir inventario"),
    }


def _inventario_command_center(*, governance_rows: list[dict[str, object]], maturity_summary: dict[str, object]) -> dict[str, object]:
    blockers = sum(int(row.get("blockers") or 0) for row in governance_rows)
    primary_row = max(governance_rows, key=lambda row: int(row.get("blockers") or 0), default={}) if governance_rows else {}
    tone = "success" if blockers == 0 else ("warning" if blockers <= 5 else "danger")
    status = "Listo para operar" if blockers == 0 else ("En atención" if blockers <= 5 else "Crítico")
    return {
        "owner": primary_row.get("owner") or "Inventario / Operación",
        "status": status,
        "tone": tone,
        "blockers": blockers,
        "next_step": maturity_summary.get("next_priority_detail") or "Continuar cierre documental del módulo.",
        "cta": maturity_summary.get("next_priority_cta") or primary_row.get("cta") or "Abrir",
        "url": maturity_summary.get("next_priority_url") or primary_row.get("url") or reverse("inventario:existencias"),
    }


def _inventario_erp_governance_rows(
    *,
    total_rows: int,
    master_blocked_count: int,
    critical_count: int,
    pending_count: int,
    healthy_count: int,
) -> list[dict[str, object]]:
    total = max(total_rows, 1)

    def _pct(done: int) -> int:
        return round((done / total) * 100) if total else 0

    master_ready = max(total_rows - master_blocked_count, 0)
    trace_ready = max(total_rows - pending_count, 0)
    inventory_ready = max(total_rows - critical_count, 0)
    release_ready = max(total_rows - max(master_blocked_count, critical_count), 0)

    return [
        {
            "front": "Maestro ERP",
            "owner": "Maestros / DG",
            "blockers": master_blocked_count,
            "completion": _pct(master_ready),
            "detail": "Unidad, clase y dato maestro deben quedar completos para liberar existencias y reorden.",
            "next_step": "Corregir artículos incompletos",
            "url": reverse("maestros:insumo_list") + f"?{urlencode({'usage_scope': 'inventory', 'enterprise_status': 'incompletos'})}",
            "cta": "Abrir maestro",
        },
        {
            "front": "Existencias y reorden",
            "owner": "Inventario / Almacén",
            "blockers": critical_count,
            "completion": _pct(inventory_ready),
            "detail": "Stock actual y punto de reorden deben sostener la lectura operativa del módulo.",
            "next_step": "Atender faltantes y reorden",
            "url": reverse("inventario:alertas"),
            "cta": "Abrir alertas",
        },
        {
            "front": "Movimientos y ajustes",
            "owner": "Inventario / Ejecución",
            "blockers": pending_count,
            "completion": _pct(trace_ready),
            "detail": "La trazabilidad documental debe quedar sin variantes activas ni referencias pendientes.",
            "next_step": "Cerrar trazabilidad documental",
            "url": reverse("inventario:movimientos"),
            "cta": "Abrir movimientos",
        },
        {
            "front": "Reabasto y entrega",
            "owner": "Compras / Producción",
            "blockers": max(master_blocked_count, critical_count),
            "completion": _pct(release_ready),
            "detail": "Inventario confiable se convierte en señal formal para compra o producción.",
            "next_step": "Emitir reabasto ERP",
            "url": reverse("inventario:existencias"),
            "cta": "Abrir existencias",
        },
    ]


def _inventario_downstream_handoff_rows(
    governance_rows: list[dict[str, object]],
    *,
    focus_label: str,
) -> list[dict[str, object]]:
    dependency_map = {
        "Maestro ERP": "Artículo maestro listo, sin brechas críticas de unidad, categoría o proveedor.",
        "Existencias y reorden": "Stock y punto de reorden confiables para la lectura operativa del día.",
        "Movimientos y ajustes": "Trazabilidad documental cerrada sin pendientes de referencia o aprobación.",
        "Reabasto y entrega": "Inventario ya listo para disparar compra, producción o surtido.",
    }
    exit_map = {
        "Maestro ERP": "Los artículos ya no bloquean consumo, ajuste ni reabasto.",
        "Existencias y reorden": "Las existencias ya sostienen alertas, cobertura y lectura de reorden.",
        "Movimientos y ajustes": "Los movimientos y ajustes ya quedan auditables y conciliados.",
        "Reabasto y entrega": "El módulo ya entrega señal confiable a compras, producción y operación.",
    }
    return [
        {
            "front": row["front"],
            "owner": row["owner"],
            "status": (
                "Controlado"
                if int(row.get("blockers") or 0) <= 0 and int(row.get("completion") or 0) >= 90
                else "En seguimiento"
                if int(row.get("completion") or 0) >= 50
                else "Con bloqueo"
            ),
            "tone": (
                "success"
                if int(row.get("blockers") or 0) <= 0 and int(row.get("completion") or 0) >= 90
                else "warning"
                if int(row.get("completion") or 0) >= 50
                else "danger"
            ),
            "blockers": int(row.get("blockers") or 0),
            "depends_on": dependency_map.get(str(row["front"]), f"Control documental previo para {focus_label.lower()}."),
            "exit_criteria": exit_map.get(str(row["front"]), "Salida operativa lista para downstream."),
            "completion": row["completion"],
            "detail": row["detail"],
            "next_step": row["next_step"],
            "url": row["url"],
            "cta": row["cta"],
        }
        for row in governance_rows
    ]


def _inventario_handoff_map(
    *,
    blocked_count: int,
    pending_count: int,
    critical_count: int,
    healthy_count: int,
) -> list[dict[str, object]]:
    return [
        {
            "label": "Maestro",
            "count": blocked_count,
            "status": "Con bloqueo" if blocked_count else "Listo",
            "detail": "Artículos ERP completos para permitir inventario, compras y costeo consistentes.",
            "tone": "warning" if blocked_count else "success",
            "url": reverse("maestros:insumo_list") + f"?{urlencode({'usage_scope': 'inventory', 'enterprise_status': 'incompletos'})}" if blocked_count else reverse("maestros:insumo_list"),
            "cta": "Corregir maestro" if blocked_count else "Abrir maestro",
        },
        {
            "label": "Movimientos",
            "count": pending_count,
            "status": "Por consolidar" if pending_count else "Controlado",
            "detail": "Entradas, salidas y consumo deben apuntar al artículo estándar sin dividir historial.",
            "tone": "warning" if pending_count else "success",
            "url": reverse("inventario:movimientos"),
            "cta": "Abrir movimientos",
        },
        {
            "label": "Ajustes",
            "count": critical_count,
            "status": "Revisión urgente" if critical_count else "Sin urgencia",
            "detail": "Diferencias críticas y faltantes deben cerrarse con evidencia y autorización.",
            "tone": "danger" if critical_count else "success",
            "url": reverse("inventario:ajustes"),
            "cta": "Abrir ajustes",
        },
        {
            "label": "Cobertura",
            "count": healthy_count,
            "status": "Suficiente" if healthy_count else "Sin cobertura sana",
            "detail": "Existencias sanas habilitan reabasto, MRP y surtido operativo sin desvíos.",
            "tone": "success" if healthy_count else "warning",
            "url": reverse("inventario:alertas") + "?nivel=ok",
            "cta": "Ver cobertura",
        },
    ]


def _inventario_critical_path_rows(chain: list[dict[str, object]]) -> list[dict[str, object]]:
    severity_order = {"danger": 0, "warning": 1, "success": 2, "primary": 3}
    ranked = sorted(
        chain,
        key=lambda item: (
            severity_order.get(str(item.get("tone") or "warning"), 9),
            -int(item.get("count") or 0),
            int(item.get("completion") or 0),
        ),
    )
    rows: list[dict[str, object]] = []
    for index, item in enumerate(ranked[:4], start=1):
        rows.append(
            {
                "rank": f"R{index}",
                "title": item.get("title", "Tramo de inventario"),
                "owner": item.get("owner", "Inventario / Operación"),
                "status": item.get("status", "Sin estado"),
                "tone": item.get("tone", "warning"),
                "count": int(item.get("count") or 0),
                "completion": int(item.get("completion") or 0),
                "depends_on": item.get("depends_on", "Inicio del flujo"),
                "dependency_status": item.get("dependency_status", "Sin dependencia registrada"),
                "detail": item.get("detail", ""),
                "next_step": item.get("next_step", "Revisar tramo"),
                "url": item.get("url", reverse("inventario:existencias")),
                "cta": item.get("cta", "Abrir"),
            }
        )
    return rows


def _inventario_executive_radar_rows(
    governance_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for row in governance_rows[:4]:
        completion = int(row.get("completion") or 0)
        blockers = int(row.get("blockers") or 0)
        if blockers <= 0 and completion >= 90:
            tone = "success"
            status = "Controlado"
            dominant_blocker = "Sin bloqueo activo"
        elif completion >= 50:
            tone = "warning"
            status = "En seguimiento"
            dominant_blocker = row.get("detail", "") or "Brecha operativa en seguimiento"
        else:
            tone = "danger"
            status = "Con bloqueo"
            dominant_blocker = row.get("detail", "") or "Bloqueo operativo abierto"
        rows.append(
            {
                "phase": row.get("front", "Frente operativo"),
                "owner": row.get("owner", "Inventario / Operación"),
                "status": status,
                "tone": tone,
                "blockers": blockers,
                "progress_pct": completion,
                "dominant_blocker": dominant_blocker,
                "depends_on": row.get("front", "Inicio del flujo"),
                "dependency_status": row.get("next_step", "Sin dependencia registrada"),
                "next_step": row.get("next_step", "Abrir frente"),
                "url": row.get("url", reverse("inventario:existencias")),
                "cta": row.get("cta", "Abrir"),
            }
        )
    return rows


def _inventario_upstream_dependency_rows(
    *,
    focus: str,
    master_blocked_count: int,
    pending_count: int,
    critical_count: int,
) -> list[dict[str, object]]:
    def _tone(blockers: int) -> str:
        return "success" if blockers <= 0 else ("danger" if blockers > 0 and focus in {"existencias", "alertas"} else "warning")

    def _completion(blockers: int) -> int:
        return 100 if blockers <= 0 else 0

    abastecimiento_detail = (
        "Existencias y alertas dependen de recepciones aplicadas y reabasto formal para sostener cobertura."
        if focus in {"existencias", "alertas"}
        else "Los movimientos y ajustes deben sostener un abastecimiento trazable antes de reflejarse en cobertura."
    )
    abastecimiento_next = (
        "Atender faltantes y convertir la señal en compra o producción."
        if focus in {"existencias", "alertas"}
        else "Cerrar recepciones, diferencias y abastecimiento antes de consolidar stock."
    )
    ledger_detail = (
        "El inventario depende de movimientos y ajustes sin variantes activas ni trazabilidad pendiente."
    )
    ledger_next = (
        "Cerrar trazabilidad del ledger y consolidar referencias del artículo."
    )
    return [
        {
            "label": "Maestro de artículos",
            "owner": "Maestros / DG",
            "status": "Listo" if master_blocked_count <= 0 else "En revisión",
            "tone": _tone(master_blocked_count),
            "blockers": master_blocked_count,
            "completion": _completion(master_blocked_count),
            "depends_on": "Unidad base, clase, proveedor y categoría",
            "exit_criteria": "Artículo maestro completo y liberado para operación.",
            "detail": "Todo stock debe depender de un artículo ERP cerrado y auditable.",
            "next_step": "Corregir artículos incompletos antes de operar inventario.",
            "url": reverse("maestros:insumo_list") + f"?{urlencode({'usage_scope': 'inventory', 'enterprise_status': 'incompletos'})}",
            "cta": "Abrir maestro",
        },
        {
            "label": "Recepciones y abastecimiento",
            "owner": "Compras / Recepción",
            "status": "Listo" if critical_count <= 0 else "En revisión",
            "tone": _tone(critical_count),
            "blockers": critical_count,
            "completion": _completion(critical_count),
            "depends_on": "Recepciones aplicadas, diferencias cerradas y reabasto documentado",
            "exit_criteria": "Abastecimiento registrado y reflejado sin brecha operativa.",
            "detail": abastecimiento_detail,
            "next_step": abastecimiento_next,
            "url": reverse("compras:recepciones"),
            "cta": "Abrir recepciones",
        },
        {
            "label": "Ledger y conciliación",
            "owner": "Inventario / Ejecución",
            "status": "Listo" if pending_count <= 0 else "En revisión",
            "tone": _tone(pending_count),
            "blockers": pending_count,
            "completion": _completion(pending_count),
            "depends_on": "Movimientos, ajustes y referencias consolidadas",
            "exit_criteria": "Kardex sin pendientes documentales ni variantes activas.",
            "detail": ledger_detail,
            "next_step": ledger_next,
            "url": reverse("inventario:movimientos"),
            "cta": "Abrir ledger",
        },
    ]


def _inventory_sales_demand_signal(rows: list[SimpleNamespace], *, lookback_days: int = 45) -> dict[str, object] | None:
    insumo_ids = [int(row.insumo.id) for row in rows if getattr(row, "insumo", None)]
    if not insumo_ids:
        return None

    since = timezone.localdate() - timedelta(days=lookback_days)
    reference_date = timezone.localdate()
    receta_ids = list(
        LineaReceta.objects.filter(
            insumo_id__in=insumo_ids,
            receta__tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        .values_list("receta_id", flat=True)
        .distinct()
    )
    if not receta_ids:
        return {
            "available": False,
            "status": "Sin vínculo comercial",
            "tone": "warning",
            "detail": "Los artículos visibles todavía no están ligados a productos con venta histórica reciente.",
            "days_count": 0,
            "recipes_count": 0,
            "insumos_count": 0,
            "units_total": Decimal("0"),
            "years_observed": 0,
            "comparable_years": 0,
            "top_products": [],
        }

    historico_recent_qs = VentaHistorica.objects.filter(receta_id__in=receta_ids, fecha__gte=since)
    historico_all_qs = VentaHistorica.objects.filter(receta_id__in=receta_ids)
    years_observed = historico_all_qs.dates("fecha", "year").count()
    comparable_years = (
        historico_all_qs.filter(fecha__month=reference_date.month).dates("fecha", "year").count()
    )
    if not historico_recent_qs.exists():
        return {
            "available": False,
            "status": "Sin ventas recientes",
            "tone": "warning",
            "detail": "No hay ventas históricas recientes para los productos que impactan estos artículos.",
            "days_count": 0,
            "recipes_count": len(set(receta_ids)),
            "insumos_count": 0,
            "units_total": Decimal("0"),
            "years_observed": years_observed,
            "comparable_years": comparable_years,
            "top_products": [],
        }

    historico_days = historico_recent_qs.values("fecha").distinct().count()
    units_total = historico_recent_qs.aggregate(total=Sum("cantidad")).get("total") or Decimal("0")
    products_count = historico_recent_qs.values("receta_id").distinct().count()
    linked_insumos_count = LineaReceta.objects.filter(
        receta_id__in=historico_recent_qs.values("receta_id").distinct(),
        insumo_id__in=insumo_ids,
    ).values("insumo_id").distinct().count()
    top_products = list(
        historico_recent_qs.values("receta__nombre").annotate(total=Sum("cantidad")).order_by("-total", "receta__nombre")[:5]
    )
    if historico_days >= 21 and comparable_years >= 3:
        status = "Demanda activa multianual"
        tone = "success"
        detail = "El inventario ya está respaldado por ventas recientes y temporadas comparables en varios años para orientar cobertura y reabasto."
    elif historico_days >= 21:
        status = "Demanda activa"
        tone = "success"
        detail = "El inventario ya está respaldado por ventas recientes suficientes para orientar cobertura y reabasto."
    elif historico_days >= 7 and years_observed >= 2:
        status = "Demanda utilizable multianual"
        tone = "warning"
        detail = "La señal comercial reciente ya tiene respaldo multianual, aunque todavía conviene validar cobertura manualmente."
    elif historico_days >= 7:
        status = "Demanda utilizable"
        tone = "warning"
        detail = "La señal comercial ya orienta el inventario, pero todavía conviene validar cobertura manualmente."
    else:
        status = "Demanda limitada"
        tone = "danger"
        detail = "La cobertura histórica reciente es corta para tomar decisiones agresivas de reabasto."

    return {
        "available": True,
        "status": status,
        "tone": tone,
        "detail": detail,
        "days_count": historico_days,
        "recipes_count": products_count,
        "insumos_count": linked_insumos_count,
        "units_total": units_total,
        "years_observed": years_observed,
        "comparable_years": comparable_years,
        "top_products": top_products,
    }


def _inventory_sales_demand_gate(signal: dict[str, object] | None) -> dict[str, object]:
    if not signal or not bool(signal.get("available")):
        return {
            "status": "Sin base comercial",
            "tone": "danger",
            "is_ready": False,
            "blockers": 1,
            "detail": "Todavía no hay señal comercial suficiente para sostener inventario y reabasto.",
            "next_step": "Valida histórico y productos ligados antes de usar esta señal como base operativa.",
        }
    tone = str(signal.get("tone") or "warning")
    years_observed = int(signal.get("years_observed") or 0)
    if tone == "success":
        return {
            "status": str(signal.get("status") or "Demanda lista"),
            "tone": "success",
            "is_ready": True,
            "blockers": 0,
            "detail": str(signal.get("detail") or ""),
            "next_step": "La señal comercial ya puede respaldar cobertura y reabasto.",
        }
    if tone == "warning" and years_observed >= 2:
        return {
            "status": str(signal.get("status") or "Demanda en revisión"),
            "tone": "warning",
            "is_ready": True,
            "blockers": 0,
            "detail": str(signal.get("detail") or ""),
            "next_step": "Usa la señal con criterio y valida cobertura antes de disparar compras agresivas.",
        }
    return {
        "status": "Demanda frágil",
        "tone": "danger",
        "is_ready": False,
        "blockers": 1,
        "detail": str(signal.get("detail") or "La base comercial todavía es corta para empujar reabasto agresivo."),
        "next_step": "Apóyate primero en maestro, consumo y validación manual antes de escalar compras.",
    }


def _inventory_commercial_priority_rows(rows: list[SimpleNamespace], *, lookback_days: int = 45, limit: int = 6) -> list[dict[str, object]]:
    insumo_ids = [int(row.insumo.id) for row in rows if getattr(row, "insumo", None)]
    if not insumo_ids:
        return []

    since = timezone.localdate() - timedelta(days=lookback_days)
    receta_map: dict[int, list[int]] = defaultdict(list)
    for receta_id, insumo_id in (
        LineaReceta.objects.filter(
            insumo_id__in=insumo_ids,
            receta__tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        .exclude(tipo_linea=LineaReceta.TIPO_SUBSECCION)
        .values_list("receta_id", "insumo_id")
    ):
        receta_map[int(insumo_id)].append(int(receta_id))

    all_receta_ids = sorted({rid for receta_ids in receta_map.values() for rid in receta_ids})
    if all_receta_ids:
        historico_map = {
            int(row["receta_id"]): Decimal(str(row["total"] or 0))
            for row in (
                VentaHistorica.objects.filter(receta_id__in=all_receta_ids, fecha__gte=since)
                .values("receta_id")
                .annotate(total=Sum("cantidad"))
            )
        }
        receta_name_map = {
            int(row["id"]): row["nombre"]
            for row in Receta.objects.filter(id__in=all_receta_ids).values("id", "nombre")
        }
    else:
        historico_map = {}
        receta_name_map = {}

    priority_rows: list[dict[str, object]] = []
    for row in rows:
        insumo = getattr(row, "insumo", None)
        if not insumo:
            continue
        canonical = canonical_insumo_by_id(int(insumo.id)) or insumo
        readiness_profile = enterprise_readiness_profile(canonical)
        linked_receta_ids = receta_map.get(int(insumo.id), [])
        historico_units = sum((historico_map.get(receta_id, Decimal("0")) for receta_id in linked_receta_ids), Decimal("0"))
        gap = max(
            Decimal(str(getattr(row, "punto_reorden", 0) or 0)) - Decimal(str(getattr(row, "stock_actual", 0) or 0)),
            Decimal("0"),
        )
        priority_score = historico_units * max(gap, Decimal("1"))
        if historico_units >= Decimal("40") and gap > 0:
            priority_label = "Alta"
            priority_tone = "danger"
        elif historico_units >= Decimal("15"):
            priority_label = "Media"
            priority_tone = "warning"
        else:
            priority_label = "Base"
            priority_tone = "primary"
        priority_rows.append(
            {
                "insumo_nombre": insumo.nombre,
                "historico_units": historico_units,
                "gap": gap,
                "stock_actual": Decimal(str(getattr(row, "stock_actual", 0) or 0)),
                "reorder_point": Decimal(str(getattr(row, "punto_reorden", 0) or 0)),
                "priority_label": priority_label,
                "priority_tone": priority_tone,
                "master_missing": readiness_profile["missing"][:2] or ["Sin faltante"],
                "recipe_names": [receta_name_map[rid] for rid in linked_receta_ids[:3] if rid in receta_name_map],
                "action_url": reverse("maestros:insumo_update", args=[canonical.id]),
                "priority_score": priority_score,
            }
        )

    priority_rows.sort(
        key=lambda item: (
            Decimal(str(item["priority_score"] or 0)),
            Decimal(str(item["historico_units"] or 0)),
            Decimal(str(item["gap"] or 0)),
        ),
        reverse=True,
    )
    return priority_rows[:limit]


def _inventory_critical_master_demand_rows(rows: list[dict[str, object]] | None, *, limit: int = 3) -> list[dict[str, object]]:
    if not rows:
        return []
    filtered = [
        row for row in rows
        if str(row.get("priority_tone") or "") == "danger"
        and any(str(item).strip().lower() != "sin faltante" for item in (row.get("master_missing") or []))
    ]
    filtered.sort(
        key=lambda item: (
            Decimal(str(item.get("historico_units") or 0)),
            Decimal(str(item.get("gap") or 0)),
        ),
        reverse=True,
    )
    return filtered[:limit]


def _inventory_daily_critical_close_focus(rows: list[dict[str, object]] | None) -> dict[str, object] | None:
    critical_rows = _inventory_critical_master_demand_rows(rows, limit=1)
    if not critical_rows:
        return None
    top_row = critical_rows[0]
    return {
        "title": "Cierre prioritario del día",
        "detail": (
            f"{top_row['insumo_nombre']} debe cerrarse primero para liberar cobertura, reorden y lectura confiable de inventario."
        ),
        "historico_units": top_row["historico_units"],
        "gap": top_row["gap"],
        "missing": top_row["master_missing"],
        "recipe_names": top_row["recipe_names"],
        "url": top_row["action_url"],
        "cta": "Cerrar artículo ahora",
        "tone": "danger",
    }


def _inventory_supply_focus_rows(rows: list[dict[str, object]] | None, *, limit: int = 6) -> list[dict[str, object]]:
    if not rows:
        return []

    prioritized = [
        row for row in rows
        if (
            Decimal(str(row.get("gap") or 0)) > 0
            or any(str(item).strip().lower() != "sin faltante" for item in (row.get("master_missing") or []))
        )
    ]
    prioritized.sort(
        key=lambda item: (
            Decimal(str(item.get("historico_units") or 0)) * max(Decimal(str(item.get("gap") or 0)), Decimal("1")),
            Decimal(str(item.get("gap") or 0)),
        ),
        reverse=True,
    )

    supply_rows: list[dict[str, object]] = []
    for row in prioritized[:limit]:
        recipe_names = list(row.get("recipe_names") or [])
        supply_rows.append(
            {
                "insumo_nombre": row.get("insumo_nombre") or "Artículo",
                "dominant_recipe_name": recipe_names[0] if recipe_names else "Sin producto dominante",
                "historico_units": Decimal(str(row.get("historico_units") or 0)),
                "stock_actual": Decimal(str(row.get("stock_actual") or 0)),
                "reorder_point": Decimal(str(row.get("reorder_point") or 0)),
                "gap": Decimal(str(row.get("gap") or 0)),
                "master_missing": list(row.get("master_missing") or []),
                "action_url": row.get("action_url") or reverse("inventario:existencias"),
                "action_label": "Asegurar artículo",
            }
        )
    return supply_rows


@login_required
def existencias(request: HttpRequest) -> HttpResponse:
    if not can_view_inventario(request.user):
        raise PermissionDenied("No tienes permisos para ver Inventario.")

    if request.method == "POST":
        if not can_manage_inventario(request.user):
            raise PermissionDenied("No tienes permisos para editar existencias.")
        action = (request.POST.get("action") or "").strip()
        if action == "update_reorder_config":
            max_diff_pct = _to_decimal(request.POST.get("reorder_max_diff_pct"), "10")
            if max_diff_pct < 0:
                messages.error(request, "El umbral máximo de desviación no puede ser negativo.")
                return redirect("inventario:existencias")
            config = InventarioConfig.get_solo()
            config.reorder_max_diff_pct = max_diff_pct
            config.save(update_fields=["reorder_max_diff_pct", "updated_at"])
            messages.success(
                request,
                f"Configuración guardada: umbral máximo manual en punto de reorden = {max_diff_pct}%.",
            )
            return redirect("inventario:existencias")

        insumo = canonical_insumo_by_id(request.POST.get("insumo_id"))
        if insumo:
            existencia, _ = ExistenciaInsumo.objects.get_or_create(insumo=insumo)
            prev_stock = existencia.stock_actual
            prev_reorden = existencia.punto_reorden
            prev_minimo = existencia.stock_minimo
            prev_maximo = existencia.stock_maximo
            prev_inv_prom = existencia.inventario_promedio
            prev_dias = existencia.dias_llegada_pedido
            prev_consumo = existencia.consumo_diario_promedio
            new_stock = _to_decimal(request.POST.get("stock_actual"), "0")
            new_minimo = _to_decimal(request.POST.get("stock_minimo"), "0")
            new_maximo = _to_decimal(request.POST.get("stock_maximo"), "0")
            new_inv_prom = _to_decimal(request.POST.get("inventario_promedio"), "0")
            new_dias = int(_to_decimal(request.POST.get("dias_llegada_pedido"), "0"))
            new_consumo = _to_decimal(request.POST.get("consumo_diario_promedio"), "0")
            reorden_recomendado = calcular_punto_reorden(
                stock_minimo=new_minimo,
                dias_llegada_pedido=new_dias,
                consumo_diario_promedio=new_consumo,
                formula=getattr(settings, "INVENTARIO_REORDER_FORMULA", FORMULA_EXCEL_LEGACY),
            )
            reorden_raw = (request.POST.get("punto_reorden") or "").strip()
            if reorden_raw:
                new_reorden = _to_decimal(reorden_raw, "0")
                reorden_auto = False
                max_diff_pct = _inventario_reorder_max_diff_pct()
                if reorden_recomendado > 0 and max_diff_pct >= 0:
                    diff_pct = (abs(new_reorden - reorden_recomendado) / reorden_recomendado) * Decimal("100")
                    if diff_pct > max_diff_pct:
                        messages.error(
                            request,
                            (
                                "El punto de reorden manual difiere demasiado del recomendado "
                                f"({diff_pct:.2f}% > {max_diff_pct:.2f}%). "
                                f"Recomendado: {reorden_recomendado}."
                            ),
                        )
                        return redirect("inventario:existencias")
            else:
                new_reorden = reorden_recomendado
                reorden_auto = True
            if (
                new_stock < 0
                or new_reorden < 0
                or new_minimo < 0
                or new_maximo < 0
                or new_inv_prom < 0
                or new_dias < 0
                or new_consumo < 0
            ):
                messages.error(request, "Los indicadores de inventario no pueden ser negativos.")
                return redirect("inventario:existencias")

            existencia.stock_actual = new_stock
            existencia.punto_reorden = new_reorden
            existencia.stock_minimo = new_minimo
            existencia.stock_maximo = new_maximo
            existencia.inventario_promedio = new_inv_prom
            existencia.dias_llegada_pedido = new_dias
            existencia.consumo_diario_promedio = new_consumo
            existencia.actualizado_en = timezone.now()
            existencia.save()
            if reorden_auto:
                messages.info(
                    request,
                    f"Punto de reorden calculado automáticamente: {new_reorden} (según fórmula activa).",
                )
            log_event(
                request.user,
                "UPDATE",
                "inventario.ExistenciaInsumo",
                existencia.id,
                {
                    "insumo_id": existencia.insumo_id,
                    "from_stock": str(prev_stock),
                    "to_stock": str(existencia.stock_actual),
                    "from_reorden": str(prev_reorden),
                    "to_reorden": str(existencia.punto_reorden),
                    "from_stock_minimo": str(prev_minimo),
                    "to_stock_minimo": str(existencia.stock_minimo),
                    "from_stock_maximo": str(prev_maximo),
                    "to_stock_maximo": str(existencia.stock_maximo),
                    "from_inventario_promedio": str(prev_inv_prom),
                    "to_inventario_promedio": str(existencia.inventario_promedio),
                    "from_dias_llegada_pedido": prev_dias,
                    "to_dias_llegada_pedido": existencia.dias_llegada_pedido,
                    "from_consumo_diario_promedio": str(prev_consumo),
                    "to_consumo_diario_promedio": str(existencia.consumo_diario_promedio),
                },
            )
        return redirect("inventario:existencias")

    existencias_rows = _canonicalized_existencias_rows(limit=200)
    formula_mode = getattr(settings, "INVENTARIO_REORDER_FORMULA", FORMULA_EXCEL_LEGACY)
    selected_focus_key = (request.GET.get("master_focus_key") or "auto").strip() or "auto"
    blocker_context = _inventory_master_blocker_context(
        existencias_rows,
        usage_scope="inventory",
        focus_summary_template=(
            "El control de inventario sigue condicionado por {name} ({missing_field})."
        ),
        row_action_detail="Completa el maestro para liberar inventario, reorden y movimientos del artículo.",
        card_action_detail="Completa el dato maestro para no bloquear ajustes, movimientos y reorden.",
        current_view_url=reverse("inventario:existencias"),
        current_query={},
        selected_focus_key=selected_focus_key,
    )
    master_blocked_count = int(blocker_context.get("master_blocker_total") or 0)

    context = {
        "existencias": existencias_rows,
        "insumos": canonicalized_insumo_selector(limit=200),
        "can_manage_inventario": can_manage_inventario(request.user),
        "reorder_formula_mode": formula_mode,
        "reorder_max_diff_pct": _inventario_reorder_max_diff_pct(),
        "formula_excel_legacy": FORMULA_EXCEL_LEGACY,
        "formula_leadtime_plus_safety": FORMULA_LEADTIME_PLUS_SAFETY,
        "enterprise_chain": _inventario_enterprise_chain(
            focus="existencias",
            total_rows=len(existencias_rows),
            blocked_count=master_blocked_count,
            critical_count=sum(1 for row in existencias_rows if Decimal(str(row.stock_actual or 0)) < Decimal(str(row.punto_reorden or 0))),
            pending_count=sum(1 for row in existencias_rows if getattr(row, "canonical_pending", False)),
        ),
        "document_stage_rows": _inventario_document_stage_rows(
            focus="existencias",
            total_rows=len(existencias_rows),
            blocked_count=master_blocked_count,
            critical_count=sum(1 for row in existencias_rows if Decimal(str(row.stock_actual or 0)) < Decimal(str(row.punto_reorden or 0))),
            pending_count=sum(1 for row in existencias_rows if getattr(row, "canonical_pending", False)),
            healthy_count=sum(1 for row in existencias_rows if Decimal(str(row.stock_actual or 0)) >= Decimal(str(row.punto_reorden or 0))),
        ),
        "operational_health_cards": _inventario_operational_health_cards(
            focus="existencias",
            total_rows=len(existencias_rows),
            blocked_count=master_blocked_count,
            critical_count=sum(1 for row in existencias_rows if Decimal(str(row.stock_actual or 0)) < Decimal(str(row.punto_reorden or 0))),
            pending_count=sum(1 for row in existencias_rows if getattr(row, "canonical_pending", False)),
            healthy_count=sum(1 for row in existencias_rows if Decimal(str(row.stock_actual or 0)) >= Decimal(str(row.punto_reorden or 0))),
        ),
        "maturity_summary": _inventario_maturity_summary(
            chain=_inventario_enterprise_chain(
                focus="existencias",
                total_rows=len(existencias_rows),
                blocked_count=master_blocked_count,
                critical_count=sum(1 for row in existencias_rows if Decimal(str(row.stock_actual or 0)) < Decimal(str(row.punto_reorden or 0))),
                pending_count=sum(1 for row in existencias_rows if getattr(row, "canonical_pending", False)),
            ),
            default_url=reverse("inventario:existencias"),
        ),
        "handoff_map": _inventario_handoff_map(
            blocked_count=master_blocked_count,
            pending_count=sum(1 for row in existencias_rows if getattr(row, "canonical_pending", False)),
            critical_count=sum(1 for row in existencias_rows if Decimal(str(row.stock_actual or 0)) < Decimal(str(row.punto_reorden or 0))),
            healthy_count=sum(1 for row in existencias_rows if Decimal(str(row.stock_actual or 0)) >= Decimal(str(row.punto_reorden or 0))),
        ),
        "release_gate_rows": [
            {
                "step": "01",
                "title": "Maestro liberado",
                "detail": "Artículos con unidad y referencia ERP listas para operar sin bloqueo del maestro.",
                "completed": max(len(existencias_rows) - master_blocked_count, 0),
                "open_count": master_blocked_count,
                "total": max(len(existencias_rows), 1),
                "tone": "success" if master_blocked_count == 0 else "warning",
                "url": reverse("inventario:existencias"),
                "cta": "Abrir existencias",
            },
            {
                "step": "02",
                "title": "Stock y reorden validados",
                "detail": "Existencias con stock por encima del punto de reorden y fórmula operativa consistente.",
                "completed": sum(
                    1
                    for row in existencias_rows
                    if Decimal(str(row.stock_actual or 0)) >= Decimal(str(row.punto_reorden or 0))
                ),
                "open_count": sum(
                    1
                    for row in existencias_rows
                    if Decimal(str(row.stock_actual or 0)) < Decimal(str(row.punto_reorden or 0))
                ),
                "total": max(len(existencias_rows), 1),
                "tone": "success"
                if not any(
                    Decimal(str(row.stock_actual or 0)) < Decimal(str(row.punto_reorden or 0))
                    for row in existencias_rows
                )
                else "warning",
                "url": reverse("inventario:alertas"),
                "cta": "Abrir alertas",
            },
            {
                "step": "03",
                "title": "Movimientos y ajustes trazables",
                "detail": "La cadena documental mantiene inventario, ajustes y reabasto con evidencia operativa.",
                "completed": sum(
                    1
                    for row in existencias_rows
                    if not getattr(row, "canonical_pending", False)
                ),
                "open_count": sum(
                    1
                    for row in existencias_rows
                    if getattr(row, "canonical_pending", False)
                ),
                "total": max(len(existencias_rows), 1),
                "tone": "success"
                if not any(getattr(row, "canonical_pending", False) for row in existencias_rows)
                else "warning",
                "url": reverse("inventario:movimientos"),
                "cta": "Abrir movimientos",
            },
        ],
        **blocker_context,
    }
    context["release_gate_completion"] = (
        round(
            (
                sum(row["completed"] for row in context["release_gate_rows"])
                / sum(row["total"] for row in context["release_gate_rows"])
            )
            * 100
        )
        if context["release_gate_rows"] and sum(row["total"] for row in context["release_gate_rows"])
        else 0
    )
    context["documental_semaphore_cards"] = [
        {
            "title": row["title"],
            "tone": row["tone"],
            "completion": round((row["completed"] / row["total"]) * 100) if row["total"] else 0,
            "completed": row["completed"],
            "open_count": row["open_count"],
            "detail": row["detail"],
            "url": row["url"],
            "cta": row["cta"],
        }
        for row in context["release_gate_rows"]
    ]
    context["workflow_stage_rows"] = [
        {
            "step": row["step"],
            "title": row["title"],
            "completion": round((row["completed"] / row["total"]) * 100) if row["total"] else 0,
            "detail": row["detail"],
            "owner": (
                "Maestros / DG"
                if row["step"] == 1
                else "Inventario / Almacén"
                if row["step"] == 2
                else "Compras / Producción"
                if row["step"] == 3
                else "Inventario / Ejecución"
            ),
            "next_step": row["cta"],
            "url": row["url"],
            "tone": row["tone"],
        }
        for row in context["release_gate_rows"]
    ]
    context["erp_governance_rows"] = _inventario_erp_governance_rows(
        total_rows=len(existencias_rows),
        master_blocked_count=master_blocked_count,
        critical_count=sum(
            1
            for row in existencias_rows
            if Decimal(str(row.stock_actual or 0)) < Decimal(str(row.punto_reorden or 0))
        ),
        pending_count=sum(1 for row in existencias_rows if getattr(row, "canonical_pending", False)),
        healthy_count=sum(
            1
            for row in existencias_rows
            if Decimal(str(row.stock_actual or 0)) >= Decimal(str(row.punto_reorden or 0))
        ),
    )
    context["erp_command_center"] = _inventario_command_center(
        governance_rows=context["erp_governance_rows"],
        maturity_summary=context["maturity_summary"],
    )
    context["downstream_handoff_rows"] = _inventario_downstream_handoff_rows(
        context["erp_governance_rows"],
        focus_label="Existencias",
    )
    context["executive_radar_rows"] = _inventario_executive_radar_rows(
        context["erp_governance_rows"],
    )
    context["critical_path_rows"] = _inventario_critical_path_rows(context["enterprise_chain"])
    context["upstream_dependency_rows"] = _inventario_upstream_dependency_rows(
        focus="existencias",
        master_blocked_count=master_blocked_count,
        pending_count=sum(1 for row in existencias_rows if getattr(row, "canonical_pending", False)),
        critical_count=sum(
            1
            for row in existencias_rows
            if Decimal(str(row.stock_actual or 0)) < Decimal(str(row.punto_reorden or 0))
        ),
    )
    context["sales_demand_signal"] = _inventory_sales_demand_signal(existencias_rows)
    context["sales_demand_gate"] = _inventory_sales_demand_gate(context["sales_demand_signal"])
    context["commercial_priority_rows"] = _inventory_commercial_priority_rows(existencias_rows)
    context["critical_master_demand_rows"] = _inventory_critical_master_demand_rows(context["commercial_priority_rows"])
    context["daily_critical_close_focus"] = _inventory_daily_critical_close_focus(context["commercial_priority_rows"])
    context["supply_focus_rows"] = _inventory_supply_focus_rows(context["commercial_priority_rows"])
    context["module_handoff_cards"] = [
        {
            "label": item["label"],
            "count": item["count"],
            "status": item["status"],
            "tone": item.get("tone", "neutral"),
            "detail": item["detail"],
            "url": item["url"],
            "cta": item["cta"],
        }
        for item in context["handoff_map"]
    ]
    return render(request, "inventario/existencias.html", context)


@login_required
def movimientos(request: HttpRequest) -> HttpResponse:
    if not can_view_inventario(request.user):
        raise PermissionDenied("No tienes permisos para ver Inventario.")

    if request.method == "POST":
        if not can_manage_inventario(request.user):
            raise PermissionDenied("No tienes permisos para registrar movimientos.")
        insumo = canonical_insumo_by_id(request.POST.get("insumo_id"))
        if insumo:
            tipo = request.POST.get("tipo") or MovimientoInventario.TIPO_ENTRADA
            if tipo == MovimientoInventario.TIPO_AJUSTE:
                messages.error(request, "El tipo AJUSTE se genera automáticamente desde la pantalla de ajustes.")
                return redirect("inventario:movimientos")

            cantidad = _to_decimal(request.POST.get("cantidad"), "0")
            if cantidad <= 0:
                messages.error(request, "La cantidad del movimiento debe ser mayor a cero.")
                return redirect("inventario:movimientos")

            existencia, _ = ExistenciaInsumo.objects.get_or_create(insumo=insumo)
            if tipo in {MovimientoInventario.TIPO_SALIDA, MovimientoInventario.TIPO_CONSUMO} and existencia.stock_actual < cantidad:
                messages.error(
                    request,
                    f"Stock insuficiente para {tipo.lower()}: disponible={existencia.stock_actual}, solicitado={cantidad}.",
                )
                return redirect("inventario:movimientos")

            movimiento = MovimientoInventario.objects.create(
                fecha=request.POST.get("fecha") or timezone.now(),
                tipo=tipo,
                insumo=insumo,
                cantidad=cantidad,
                referencia=request.POST.get("referencia", "").strip(),
            )
            _apply_movimiento(movimiento)
            log_event(
                request.user,
                "CREATE",
                "inventario.MovimientoInventario",
                movimiento.id,
                {
                    "tipo": movimiento.tipo,
                    "insumo_id": movimiento.insumo_id,
                    "cantidad": str(movimiento.cantidad),
                    "referencia": movimiento.referencia,
                },
            )
        return redirect("inventario:movimientos")

    insumo_options = _canonicalized_insumo_stock_options(limit=200)
    movimiento_rows = list(MovimientoInventario.objects.select_related("insumo")[:100])
    for movimiento in movimiento_rows:
        enterprise_profile = enterprise_readiness_profile(movimiento.insumo)
        canonical = canonical_insumo_by_id(movimiento.insumo_id) or movimiento.insumo
        movimiento.enterprise_profile = enterprise_profile
        movimiento.enterprise_missing = enterprise_profile["missing"]
        movimiento.enterprise_status = enterprise_profile["readiness_label"]
        movimiento.enterprise_usage_label = (
            "Producción interna"
            if movimiento.insumo.tipo_item == Insumo.TIPO_INTERNO
            else "Empaque final"
            if movimiento.insumo.tipo_item == Insumo.TIPO_EMPAQUE
            else "Compra directa"
        )
        primary_missing = enterprise_profile["missing"][0] if enterprise_profile["missing"] else ""
        missing_field = (
            "unidad"
            if primary_missing == "unidad base"
            else "proveedor"
            if primary_missing == "proveedor principal"
            else "categoria"
            if primary_missing == "categoría"
            else "codigo_point"
            if primary_missing == "código comercial"
            else None
        )
        movement_query = {"insumo_id": movimiento.insumo_id, "usage_scope": "inventory"}
        if missing_field:
            movement_query["missing_field"] = missing_field
        movimiento.enterprise_edit_url = reverse("maestros:insumo_update", args=[movimiento.insumo_id])
        movimiento.enterprise_list_url = reverse("maestros:insumo_list") + f"?{urlencode(movement_query)}"
        movimiento.canonical_pending = canonical is not None and canonical.id != movimiento.insumo_id
        movimiento.canonical_pending_label = "Consolidación pendiente" if movimiento.canonical_pending else ""
        movimiento.canonical_list_url = (
            reverse("maestros:insumo_list")
            + f"?{urlencode({'usage_scope': 'inventory', 'canonical_status': 'variantes', 'q': canonical.nombre if canonical else movimiento.insumo.nombre})}"
        )

    selected_focus_key = (request.GET.get("master_focus_key") or "auto").strip() or "auto"
    blocker_context = _inventory_master_blocker_context(
        _canonicalized_existencias_rows(limit=200),
        usage_scope="inventory",
        focus_summary_template=(
            "La captura de movimientos sigue condicionada por {name} ({missing_field})."
        ),
        row_action_detail="Completa el maestro para registrar entradas, salidas y consumos sin bloquear trazabilidad.",
        card_action_detail="Completa el dato maestro para no bloquear movimientos, trazabilidad y control operativo.",
        current_view_url=reverse("inventario:movimientos"),
        current_query={},
        selected_focus_key=selected_focus_key,
    )
    master_blocked_count = int(blocker_context.get("master_blocker_total") or 0)

    context = {
        "movimientos": movimiento_rows,
        "insumo_options": insumo_options,
        "tipo_choices": [
            (value, label)
            for value, label in MovimientoInventario.TIPO_CHOICES
            if value != MovimientoInventario.TIPO_AJUSTE
        ],
        "can_manage_inventario": can_manage_inventario(request.user),
        "enterprise_chain": _inventario_enterprise_chain(
            focus="movimientos",
            total_rows=len(movimiento_rows),
            blocked_count=master_blocked_count,
            pending_count=sum(1 for row in movimiento_rows if getattr(row, "canonical_pending", False)),
        ),
        "document_stage_rows": _inventario_document_stage_rows(
            focus="movimientos",
            total_rows=len(movimiento_rows),
            blocked_count=master_blocked_count,
            pending_count=sum(1 for row in movimiento_rows if getattr(row, "canonical_pending", False)),
        ),
        "operational_health_cards": _inventario_operational_health_cards(
            focus="movimientos",
            total_rows=len(movimiento_rows),
            blocked_count=master_blocked_count,
            pending_count=sum(1 for row in movimiento_rows if getattr(row, "canonical_pending", False)),
        ),
        "maturity_summary": _inventario_maturity_summary(
            chain=_inventario_enterprise_chain(
                focus="movimientos",
                total_rows=len(movimiento_rows),
                blocked_count=master_blocked_count,
                pending_count=sum(1 for row in movimiento_rows if getattr(row, "canonical_pending", False)),
            ),
            default_url=reverse("inventario:movimientos"),
        ),
        "handoff_map": _inventario_handoff_map(
            blocked_count=master_blocked_count,
            pending_count=sum(1 for row in movimiento_rows if getattr(row, "canonical_pending", False)),
            critical_count=0,
            healthy_count=max(len(movimiento_rows) - sum(1 for row in movimiento_rows if getattr(row, "canonical_pending", False)), 0),
        ),
        **blocker_context,
    }
    context["erp_governance_rows"] = _inventario_erp_governance_rows(
        total_rows=len(movimiento_rows),
        master_blocked_count=master_blocked_count,
        critical_count=0,
        pending_count=sum(1 for row in movimiento_rows if getattr(row, "canonical_pending", False)),
        healthy_count=max(len(movimiento_rows) - sum(1 for row in movimiento_rows if getattr(row, "canonical_pending", False)), 0),
    )
    context["erp_command_center"] = _inventario_command_center(
        governance_rows=context["erp_governance_rows"],
        maturity_summary=context["maturity_summary"],
    )
    context["downstream_handoff_rows"] = _inventario_downstream_handoff_rows(
        context["erp_governance_rows"],
        focus_label="Movimientos",
    )
    context["executive_radar_rows"] = _inventario_executive_radar_rows(
        context["erp_governance_rows"],
    )
    context["critical_path_rows"] = _inventario_critical_path_rows(context["enterprise_chain"])
    context["upstream_dependency_rows"] = _inventario_upstream_dependency_rows(
        focus="movimientos",
        master_blocked_count=master_blocked_count,
        pending_count=sum(1 for row in movimiento_rows if getattr(row, "canonical_pending", False)),
        critical_count=0,
    )
    return render(request, "inventario/movimientos.html", context)


@login_required
def ajustes(request: HttpRequest) -> HttpResponse:
    if not can_view_inventario(request.user):
        raise PermissionDenied("No tienes permisos para ver Inventario.")

    can_approve_ajustes = _can_approve_ajustes(request.user)

    if request.method == "POST":
        action = (request.POST.get("action") or "create").strip().lower()

        if action in {"approve", "reject", "apply"}:
            if not can_approve_ajustes:
                raise PermissionDenied("No tienes permisos para aprobar/rechazar ajustes.")
            ajuste_id = request.POST.get("ajuste_id")
            ajuste = AjusteInventario.objects.select_related("insumo").filter(pk=ajuste_id).first()
            if not ajuste:
                messages.error(request, "No se encontró el ajuste seleccionado.")
                return redirect("inventario:ajustes")

            comentario = (request.POST.get("comentario_revision") or "").strip()[:255]
            if action == "reject":
                if ajuste.estatus == AjusteInventario.STATUS_APLICADO:
                    messages.error(request, "No se puede rechazar un ajuste ya aplicado.")
                    return redirect("inventario:ajustes")
                ajuste.estatus = AjusteInventario.STATUS_RECHAZADO
                ajuste.aprobado_por = request.user
                ajuste.aprobado_en = timezone.now()
                ajuste.aplicado_en = None
                ajuste.comentario_revision = comentario
                ajuste.save(
                    update_fields=[
                        "estatus",
                        "aprobado_por",
                        "aprobado_en",
                        "aplicado_en",
                        "comentario_revision",
                    ]
                )
                log_event(
                    request.user,
                    "REJECT",
                    "inventario.AjusteInventario",
                    ajuste.id,
                    {"folio": ajuste.folio, "estatus": ajuste.estatus, "comentario_revision": comentario},
                )
                messages.success(request, f"Ajuste {ajuste.folio} rechazado.")
                return redirect("inventario:ajustes")

            if ajuste.estatus == AjusteInventario.STATUS_APLICADO:
                messages.info(request, f"El ajuste {ajuste.folio} ya estaba aplicado.")
                return redirect("inventario:ajustes")
            if ajuste.estatus == AjusteInventario.STATUS_RECHAZADO:
                messages.error(request, f"El ajuste {ajuste.folio} fue rechazado y no puede aplicarse.")
                return redirect("inventario:ajustes")

            _apply_ajuste(ajuste, request.user, comentario=comentario)
            messages.success(request, f"Ajuste {ajuste.folio} aprobado y aplicado.")
            return redirect("inventario:ajustes")

        if not can_manage_inventario(request.user):
            raise PermissionDenied("No tienes permisos para registrar ajustes.")
        insumo = canonical_insumo_by_id(request.POST.get("insumo_id"))
        if insumo:
            cantidad_sistema = _to_decimal(request.POST.get("cantidad_sistema"), "0")
            cantidad_fisica = _to_decimal(request.POST.get("cantidad_fisica"), "0")
            if cantidad_sistema < 0 or cantidad_fisica < 0:
                messages.error(request, "Las cantidades del ajuste no pueden ser negativas.")
                return redirect("inventario:ajustes")

            ajuste = AjusteInventario.objects.create(
                insumo=insumo,
                cantidad_sistema=cantidad_sistema,
                cantidad_fisica=cantidad_fisica,
                motivo=request.POST.get("motivo", "").strip() or "Sin motivo",
                estatus=AjusteInventario.STATUS_PENDIENTE,
                solicitado_por=request.user if request.user.is_authenticated else None,
            )
            log_event(
                request.user,
                "CREATE",
                "inventario.AjusteInventario",
                ajuste.id,
                {
                    "folio": ajuste.folio,
                    "insumo_id": ajuste.insumo_id,
                    "cantidad_sistema": str(ajuste.cantidad_sistema),
                    "cantidad_fisica": str(ajuste.cantidad_fisica),
                    "estatus": ajuste.estatus,
                    "solicitado_por": request.user.username if request.user.is_authenticated else "",
                },
            )

            if request.POST.get("create_and_apply") == "1":
                if can_approve_ajustes:
                    comentario = (request.POST.get("comentario_revision") or "").strip()[:255]
                    _apply_ajuste(ajuste, request.user, comentario=comentario)
                    messages.success(request, f"Ajuste {ajuste.folio} creado y aplicado.")
                else:
                    messages.info(request, f"Ajuste {ajuste.folio} creado en pendiente (requiere aprobación ADMIN).")
            else:
                messages.success(request, f"Ajuste {ajuste.folio} creado en pendiente.")
        return redirect("inventario:ajustes")

    ajustes_qs = AjusteInventario.objects.select_related("insumo", "solicitado_por", "aprobado_por")[:150]
    ajustes_rows = []
    for ajuste in ajustes_qs:
        enterprise_profile = enterprise_readiness_profile(ajuste.insumo)
        canonical = canonical_insumo_by_id(ajuste.insumo_id) or ajuste.insumo
        ajuste.enterprise_profile = enterprise_profile
        ajuste.enterprise_missing = enterprise_profile["missing"]
        ajuste.enterprise_status = enterprise_profile["readiness_label"]
        ajuste.enterprise_usage_label = (
            "Producción interna"
            if ajuste.insumo.tipo_item == Insumo.TIPO_INTERNO
            else "Empaque final"
            if ajuste.insumo.tipo_item == Insumo.TIPO_EMPAQUE
            else "Compra directa"
        )
        primary_missing = enterprise_profile["missing"][0] if enterprise_profile["missing"] else ""
        missing_field = (
            "unidad"
            if primary_missing == "unidad base"
            else "proveedor"
            if primary_missing == "proveedor principal"
            else "categoria"
            if primary_missing == "categoría"
            else "codigo_point"
            if primary_missing == "código comercial"
            else None
        )
        ajuste_query = {"insumo_id": ajuste.insumo_id, "usage_scope": "inventory"}
        if missing_field:
            ajuste_query["missing_field"] = missing_field
        ajuste.enterprise_edit_url = reverse("maestros:insumo_update", args=[ajuste.insumo_id])
        ajuste.enterprise_list_url = reverse("maestros:insumo_list") + f"?{urlencode(ajuste_query)}"
        ajuste.canonical_pending = canonical is not None and canonical.id != ajuste.insumo_id
        ajuste.canonical_pending_label = "Consolidación pendiente" if ajuste.canonical_pending else ""
        ajuste.canonical_list_url = (
            reverse("maestros:insumo_list")
            + f"?{urlencode({'usage_scope': 'inventory', 'canonical_status': 'variantes', 'q': canonical.nombre if canonical else ajuste.insumo.nombre})}"
        )
        ajustes_rows.append(ajuste)

    selected_focus_key = (request.GET.get("master_focus_key") or "auto").strip() or "auto"
    blocker_context = _inventory_master_blocker_context(
        _canonicalized_existencias_rows(limit=200),
        usage_scope="inventory",
        focus_summary_template=(
            "La operación de ajustes sigue condicionada por {name} ({missing_field})."
        ),
        row_action_detail="Completa el maestro para aprobar, aplicar y auditar ajustes sin bloqueos del artículo.",
        card_action_detail="Completa el dato maestro para no bloquear ajustes, trazabilidad y control de diferencias.",
        current_view_url=reverse("inventario:ajustes"),
        current_query={},
        selected_focus_key=selected_focus_key,
    )
    master_blocked_count = int(blocker_context.get("master_blocker_total") or 0)

    context = {
        "ajustes": ajustes_rows,
        "insumos": canonicalized_insumo_selector(limit=200),
        "status_choices": AjusteInventario.STATUS_CHOICES,
        "can_manage_inventario": can_manage_inventario(request.user),
        "can_approve_ajustes": can_approve_ajustes,
        "enterprise_chain": _inventario_enterprise_chain(
            focus="ajustes",
            total_rows=len(ajustes_rows),
            blocked_count=master_blocked_count,
            pending_count=sum(1 for row in ajustes_rows if row.estatus == AjusteInventario.STATUS_PENDIENTE),
        ),
        "document_stage_rows": _inventario_document_stage_rows(
            focus="ajustes",
            total_rows=len(ajustes_rows),
            blocked_count=master_blocked_count,
            pending_count=sum(1 for row in ajustes_rows if row.estatus == AjusteInventario.STATUS_PENDIENTE),
        ),
        "operational_health_cards": _inventario_operational_health_cards(
            focus="ajustes",
            total_rows=len(ajustes_rows),
            blocked_count=master_blocked_count,
            pending_count=sum(1 for row in ajustes_rows if row.estatus == AjusteInventario.STATUS_PENDIENTE),
        ),
        "maturity_summary": _inventario_maturity_summary(
            chain=_inventario_enterprise_chain(
                focus="ajustes",
                total_rows=len(ajustes_rows),
                blocked_count=master_blocked_count,
                pending_count=sum(1 for row in ajustes_rows if row.estatus == AjusteInventario.STATUS_PENDIENTE),
            ),
            default_url=reverse("inventario:ajustes"),
        ),
        "handoff_map": _inventario_handoff_map(
            blocked_count=master_blocked_count,
            pending_count=sum(1 for row in ajustes_rows if row.estatus == AjusteInventario.STATUS_PENDIENTE),
            critical_count=sum(1 for row in ajustes_rows if row.estatus == AjusteInventario.STATUS_RECHAZADO),
            healthy_count=sum(1 for row in ajustes_rows if row.estatus == AjusteInventario.STATUS_APLICADO),
        ),
        **blocker_context,
    }
    context["erp_governance_rows"] = _inventario_erp_governance_rows(
        total_rows=len(ajustes_rows),
        master_blocked_count=master_blocked_count,
        critical_count=sum(1 for row in ajustes_rows if row.estatus == AjusteInventario.STATUS_RECHAZADO),
        pending_count=sum(1 for row in ajustes_rows if row.estatus == AjusteInventario.STATUS_PENDIENTE),
        healthy_count=sum(1 for row in ajustes_rows if row.estatus == AjusteInventario.STATUS_APLICADO),
    )
    context["erp_command_center"] = _inventario_command_center(
        governance_rows=context["erp_governance_rows"],
        maturity_summary=context["maturity_summary"],
    )
    context["downstream_handoff_rows"] = _inventario_downstream_handoff_rows(
        context["erp_governance_rows"],
        focus_label="Ajustes",
    )
    context["executive_radar_rows"] = _inventario_executive_radar_rows(
        context["erp_governance_rows"],
    )
    context["critical_path_rows"] = _inventario_critical_path_rows(context["enterprise_chain"])
    context["upstream_dependency_rows"] = _inventario_upstream_dependency_rows(
        focus="ajustes",
        master_blocked_count=master_blocked_count,
        pending_count=sum(1 for row in ajustes_rows if row.estatus == AjusteInventario.STATUS_PENDIENTE),
        critical_count=sum(1 for row in ajustes_rows if row.estatus == AjusteInventario.STATUS_RECHAZADO),
    )
    return render(request, "inventario/ajustes.html", context)


@login_required
def alertas(request: HttpRequest) -> HttpResponse:
    if not can_view_inventario(request.user):
        raise PermissionDenied("No tienes permisos para ver Inventario.")

    nivel = (request.GET.get("nivel") or "alerta").lower()
    valid_levels = {"alerta", "critico", "bajo", "ok", "all"}
    if nivel not in valid_levels:
        nivel = "alerta"

    existencias = _canonicalized_existencias_rows(limit=500)

    rows = []
    criticos_count = 0
    bajo_reorden_count = 0
    ok_count = 0

    for e in existencias:
        stock = Decimal(str(e.stock_actual or 0))
        reorden = Decimal(str(e.punto_reorden or 0))
        diferencia = stock - reorden
        if stock <= 0:
            nivel_row = "critico"
            etiqueta = "Sin stock"
            criticos_count += 1
        elif stock < reorden:
            nivel_row = "bajo"
            etiqueta = "Bajo reorden"
            bajo_reorden_count += 1
        else:
            nivel_row = "ok"
            etiqueta = "Stock suficiente"
            ok_count += 1

        include = False
        if nivel == "all":
            include = True
        elif nivel == "alerta":
            include = nivel_row in {"critico", "bajo"}
        elif nivel == nivel_row:
            include = True

        if include:
            e.alerta_nivel = nivel_row
            e.alerta_etiqueta = etiqueta
            e.alerta_diferencia = diferencia
            rows.append(e)

    selected_focus_key = (request.GET.get("master_focus_key") or "auto").strip() or "auto"
    blocker_context = _inventory_master_blocker_context(
        rows if nivel != "all" else existencias,
        usage_scope="inventory",
        focus_summary_template=(
            "La priorización de alertas sigue condicionada por {name} ({missing_field})."
        ),
        row_action_detail="Completa el maestro para liberar alertas, reorden y priorización del artículo.",
        card_action_detail="Completa el dato maestro para no bloquear alertas, reorden y priorización.",
        current_view_url=reverse("inventario:alertas"),
        current_query={"nivel": nivel} if nivel != "alerta" else {},
        selected_focus_key=selected_focus_key,
    )
    master_blocked_count = int(blocker_context.get("master_blocker_total") or 0)

    context = {
        "rows": rows,
        "nivel": nivel,
        "criticos_count": criticos_count,
        "bajo_reorden_count": bajo_reorden_count,
        "ok_count": ok_count,
        "total_count": len(existencias),
        "enterprise_chain": _inventario_enterprise_chain(
            focus="alertas",
            total_rows=len(rows),
            blocked_count=master_blocked_count,
            critical_count=criticos_count,
            pending_count=sum(1 for row in rows if getattr(row, "canonical_pending", False)),
        ),
        "document_stage_rows": _inventario_document_stage_rows(
            focus="alertas",
            total_rows=len(rows),
            blocked_count=master_blocked_count,
            critical_count=criticos_count,
            pending_count=sum(1 for row in rows if getattr(row, "canonical_pending", False)),
            healthy_count=ok_count,
        ),
        "operational_health_cards": _inventario_operational_health_cards(
            focus="alertas",
            total_rows=len(rows),
            blocked_count=master_blocked_count,
            critical_count=criticos_count,
            pending_count=sum(1 for row in rows if getattr(row, "canonical_pending", False)),
            healthy_count=ok_count,
        ),
        "maturity_summary": _inventario_maturity_summary(
            chain=_inventario_enterprise_chain(
                focus="alertas",
                total_rows=len(rows),
                blocked_count=master_blocked_count,
                critical_count=criticos_count,
                pending_count=sum(1 for row in rows if getattr(row, "canonical_pending", False)),
            ),
            default_url=reverse("inventario:alertas"),
        ),
        "handoff_map": _inventario_handoff_map(
            blocked_count=master_blocked_count,
            pending_count=sum(1 for row in rows if getattr(row, "canonical_pending", False)),
            critical_count=criticos_count,
            healthy_count=ok_count,
        ),
        **blocker_context,
    }
    context["erp_governance_rows"] = _inventario_erp_governance_rows(
        total_rows=len(rows),
        master_blocked_count=master_blocked_count,
        critical_count=criticos_count,
        pending_count=sum(1 for row in rows if getattr(row, "canonical_pending", False)),
        healthy_count=ok_count,
    )
    context["erp_command_center"] = _inventario_command_center(
        governance_rows=context["erp_governance_rows"],
        maturity_summary=context["maturity_summary"],
    )
    context["downstream_handoff_rows"] = _inventario_downstream_handoff_rows(
        context["erp_governance_rows"],
        focus_label="Alertas",
    )
    context["executive_radar_rows"] = _inventario_executive_radar_rows(
        context["erp_governance_rows"],
    )
    context["critical_path_rows"] = _inventario_critical_path_rows(context["enterprise_chain"])
    context["upstream_dependency_rows"] = _inventario_upstream_dependency_rows(
        focus="alertas",
        master_blocked_count=master_blocked_count,
        pending_count=sum(1 for row in rows if getattr(row, "canonical_pending", False)),
        critical_count=criticos_count,
    )
    context["sales_demand_signal"] = _inventory_sales_demand_signal(rows if rows else existencias)
    context["sales_demand_gate"] = _inventory_sales_demand_gate(context["sales_demand_signal"])
    context["commercial_priority_rows"] = _inventory_commercial_priority_rows(rows if rows else existencias)
    context["critical_master_demand_rows"] = _inventory_critical_master_demand_rows(context["commercial_priority_rows"])
    context["daily_critical_close_focus"] = _inventory_daily_critical_close_focus(context["commercial_priority_rows"])
    context["supply_focus_rows"] = _inventory_supply_focus_rows(context["commercial_priority_rows"])
    return render(request, "inventario/alertas.html", context)
