"""
Diagnostica nombres sin match del Excel de almacén y sugiere aliases.

Uso:
    python manage.py sugerir_aliases_excel_almacen <ruta_excel>
    python manage.py sugerir_aliases_excel_almacen <ruta_excel> --crear-aliases --umbral 85
    python manage.py sugerir_aliases_excel_almacen <ruta_excel> --csv logs/sugerencias.csv

Flujo:
  1. Lee la hoja INVENTARIO del Excel
  2. Por cada nombre, intenta match exacto normalizado
  3. Para nombres sin match exacto: fuzzy match con rapidfuzz (ratio)
  4. Muestra tabla con resultado y score
  5. Con --crear-aliases: crea InsumoAlias para matches >= umbral
"""
from __future__ import annotations

import csv
import unicodedata
import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

try:
    from rapidfuzz import fuzz, process as rf_process
    HAS_RAPIDFUZZ = True
except ImportError:
    HAS_RAPIDFUZZ = False


SKIP_NAMES = {
    "", "ALMACEN 1", "ALMACEN 2", "ALMACEN LIMPIEZA", "ALMACEN DE LIMPIEZA",
    "ALMACEN  CASA 2", "ALMACEN DE VELAS", "ALMACEN CASA 1", "ALMACÈN 1",
    "CUARTO FRIO", "CUARTO DE LIMPIEZA", "NONE",
}


def _norm(name: str) -> str:
    try:
        from recetas.utils.normalizacion import normalizar_nombre
        return normalizar_nombre(name or "")
    except Exception:
        s = unicodedata.normalize("NFKD", (name or "").upper())
        s = "".join(c for c in s if not unicodedata.combining(c))
        return re.sub(r"\s+", " ", s).strip()


class Command(BaseCommand):
    help = "Diagnostica y sugiere aliases para nombres sin match del Excel de almacén"

    def add_arguments(self, parser):
        parser.add_argument("excel", help="Ruta al archivo .xlsx")
        parser.add_argument(
            "--umbral", type=float, default=80.0,
            help="Score mínimo (0-100) para considerar un match fuzzy válido (default: 80)"
        )
        parser.add_argument(
            "--crear-aliases", action="store_true",
            help="Crear InsumoAlias para matches con score >= umbral"
        )
        parser.add_argument(
            "--csv", default="",
            help="Ruta de CSV de salida con todas las sugerencias"
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError("openpyxl no está instalado")

        path = Path(options["excel"])
        if not path.exists():
            raise CommandError(f"Archivo no encontrado: {path}")

        umbral = options["umbral"]
        crear = options["crear_aliases"]
        csv_path = options["csv"]

        from maestros.models import Insumo, InsumoAlias

        insumos = list(Insumo.objects.filter(activo=True).select_related("unidad_base"))
        aliases = list(InsumoAlias.objects.select_related("insumo"))

        nombre_map: dict[str, Insumo] = {_norm(i.nombre): i for i in insumos}
        alias_map: dict[str, Insumo] = {_norm(a.nombre): a.insumo for a in aliases}

        # Fuzzy candidates: list of (norm_name, insumo)
        fuzzy_candidates = [(norm, ins) for norm, ins in nombre_map.items()]
        fuzzy_norms = [n for n, _ in fuzzy_candidates]

        wb = openpyxl.load_workbook(str(path), data_only=True)
        if "INVENTARIO" not in wb.sheetnames:
            raise CommandError("No existe la hoja INVENTARIO en el archivo")

        ws = wb["INVENTARIO"]
        nombres_excel: list[str] = []
        seen: set[str] = set()

        for r in range(7, ws.max_row + 1):
            val = ws.cell(r, 2).value
            if not val:
                continue
            nombre = str(val).strip()
            if nombre.upper() in SKIP_NAMES or nombre in seen:
                continue
            seen.add(nombre)
            nombres_excel.append(nombre)

        self.stdout.write(f"Total nombres únicos en Excel: {len(nombres_excel)}")

        matched: list[dict] = []
        unmatched: list[dict] = []

        for nombre in nombres_excel:
            norm = _norm(nombre)

            # 1) exact match
            insumo = nombre_map.get(norm) or alias_map.get(norm)
            if insumo:
                matched.append({"excel": nombre, "erp": insumo.nombre, "score": 100, "insumo_id": insumo.id})
                continue

            # 2) partial: strip trailing parentheses
            base = norm.split("(")[0].strip()
            if base and base in nombre_map:
                insumo = nombre_map[base]
                matched.append({"excel": nombre, "erp": insumo.nombre, "score": 99, "insumo_id": insumo.id})
                continue

            # 3) fuzzy
            if HAS_RAPIDFUZZ and fuzzy_norms:
                best = rf_process.extractOne(norm, fuzzy_norms, scorer=fuzz.token_sort_ratio)
                if best and best[1] >= umbral:
                    idx = fuzzy_norms.index(best[0])
                    sugerido = fuzzy_candidates[idx][1]
                    unmatched.append({
                        "excel": nombre,
                        "sugerido_erp": sugerido.nombre,
                        "score": round(best[1], 1),
                        "insumo_id": sugerido.id,
                        "accion": "SUGERIDO",
                    })
                else:
                    unmatched.append({
                        "excel": nombre,
                        "sugerido_erp": best[0] if best else "—",
                        "score": round(best[1], 1) if best else 0,
                        "insumo_id": None,
                        "accion": "SIN_MATCH",
                    })
            else:
                unmatched.append({
                    "excel": nombre, "sugerido_erp": "—", "score": 0,
                    "insumo_id": None, "accion": "SIN_MATCH",
                })

        self.stdout.write(self.style.SUCCESS(f"Con match exacto : {len(matched)}"))
        self.stdout.write(self.style.WARNING(f"Sin match exacto : {len(unmatched)}"))

        # Separate sugeridos vs sin_match
        sugeridos = [u for u in unmatched if u["accion"] == "SUGERIDO"]
        sin_match = [u for u in unmatched if u["accion"] == "SIN_MATCH"]

        self.stdout.write(f"  Fuzzy sugerido (>= {umbral}): {len(sugeridos)}")
        self.stdout.write(f"  Sin match alguno            : {len(sin_match)}")

        # Print fuzzy suggestions table
        if sugeridos:
            self.stdout.write("\n--- Sugerencias fuzzy ---")
            for s in sorted(sugeridos, key=lambda x: -x["score"]):
                self.stdout.write(
                    f"  [{s['score']:5.1f}]  {s['excel'][:40]:<40}  →  {s['sugerido_erp']}"
                )

        if sin_match:
            self.stdout.write("\n--- Sin match (revisar manualmente) ---")
            for s in sin_match:
                self.stdout.write(f"  {s['excel']}")

        # Create aliases
        aliases_created = 0
        if crear:
            existing_alias_norms = {_norm(a.nombre) for a in InsumoAlias.objects.all()}
            for s in sugeridos:
                if s["insumo_id"] is None:
                    continue
                alias_norm = _norm(s["excel"])
                if alias_norm in existing_alias_norms or alias_norm in nombre_map:
                    continue
                try:
                    insumo = Insumo.objects.get(pk=s["insumo_id"])
                    InsumoAlias.objects.create(insumo=insumo, nombre=s["excel"])
                    existing_alias_norms.add(alias_norm)
                    aliases_created += 1
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f"  Error creando alias '{s['excel']}': {e}"))
            self.stdout.write(self.style.SUCCESS(f"\nAliases creados: {aliases_created}"))

        # CSV export
        if csv_path:
            out = Path(csv_path)
            out.parent.mkdir(parents=True, exist_ok=True)
            with out.open("w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=["excel", "erp_match", "score", "accion", "insumo_id"],
                )
                writer.writeheader()
                for m in matched:
                    writer.writerow({
                        "excel": m["excel"], "erp_match": m["erp"],
                        "score": m["score"], "accion": "MATCH", "insumo_id": m["insumo_id"],
                    })
                for u in unmatched:
                    writer.writerow({
                        "excel": u["excel"], "erp_match": u["sugerido_erp"],
                        "score": u["score"], "accion": u["accion"], "insumo_id": u["insumo_id"] or "",
                    })
            self.stdout.write(f"\nCSV guardado en: {out}")
