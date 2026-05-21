"""
Importa el Excel de Control de Inventario Almacén.

Uso:
    python manage.py importar_excel_almacen <ruta_excel> [--periodo 2026-05] [--dry-run]

Sheets procesadas:
  INVENTARIO  → ExistenciaInsumo (stock_actual, stock_min/max, punto_reorden,
                consumo_diario, dias_llegada, almacen)
  ENTRADAS    → MovimientoInventario ENTRADA por día del período
  SALIDAS     → MovimientoInventario SALIDA por día del período
"""
from __future__ import annotations

import hashlib
from datetime import date, datetime, timezone as dt_timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore


ALMACEN_MAP = {
    "ALMACÈN 1": "ALMACEN_1",
    "ALMACEN 1": "ALMACEN_1",
    "ALMACEN CASA 1": "ALMACEN_CASA_1",
    "ALMACEN  CASA 2": "ALMACEN_CASA_2",
    "ALMACEN CASA 2": "ALMACEN_CASA_2",
    "CUARTO DE LIMPIEZA": "LIMPIEZA",
    "ALMACEN DE LIMPIEZA": "LIMPIEZA",
    "CUARTO FRIO": "CUARTO_FRIO",
    "ALMACEN DE VELAS": "VELAS",
}

SKIP_NAMES = {
    "", "ALMACEN 1", "ALMACEN 2", "ALMACEN LIMPIEZA", "ALMACEN DE LIMPIEZA",
    "ALMACEN  CASA 2", "ALMACEN DE VELAS", "ALMACEN CASA 1", "ALMACÈN 1",
    "CUARTO FRIO", "CUARTO DE LIMPIEZA", "NONE",
}


def _d(val) -> Decimal:
    try:
        return Decimal(str(val or 0))
    except InvalidOperation:
        return Decimal("0")


def _norm(name: str) -> str:
    try:
        from recetas.utils.normalizacion import normalizar_nombre
        return normalizar_nombre(name or "")
    except Exception:
        import unicodedata, re
        s = unicodedata.normalize("NFKD", (name or "").upper())
        s = "".join(c for c in s if not unicodedata.combining(c))
        return re.sub(r"\s+", " ", s).strip()


def _match_insumo(nombre: str, nombre_map: dict, alias_map: dict):
    norm = _norm(nombre)
    if norm in nombre_map:
        return nombre_map[norm]
    if norm in alias_map:
        return alias_map[norm]
    # Try without trailing parentheses / units
    base = norm.split("(")[0].strip()
    if base and base in nombre_map:
        return nombre_map[base]
    return None


class Command(BaseCommand):
    help = "Importa Excel de Control de Inventario Almacén"

    def add_arguments(self, parser):
        parser.add_argument("excel", help="Ruta al archivo .xlsx")
        parser.add_argument("--periodo", default="", help="Período YYYY-MM (ej: 2026-05)")
        parser.add_argument("--dry-run", action="store_true", help="Solo reportar, no guardar")

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError("openpyxl no está instalado")

        path = Path(options["excel"])
        if not path.exists():
            raise CommandError(f"Archivo no encontrado: {path}")

        dry_run = options["dry_run"]
        periodo_str = options["periodo"]

        # Detect period from filename if not provided
        if not periodo_str:
            import re
            m = re.search(r"(20\d\d)[_\-\s]*(\d{2})", path.stem)
            if m:
                periodo_str = f"{m.group(1)}-{m.group(2)}"
            else:
                periodo_str = timezone.localdate().strftime("%Y-%m")

        try:
            year, month = int(periodo_str[:4]), int(periodo_str[5:7])
        except (ValueError, IndexError):
            raise CommandError(f"Período inválido: {periodo_str}. Use YYYY-MM")

        self.stdout.write(f"Período: {year}-{month:02d}  |  dry_run={dry_run}")

        from maestros.models import Insumo, InsumoAlias
        from inventario.models import ExistenciaInsumo, MovimientoInventario

        # Build lookup maps
        nombre_map = {_norm(i.nombre): i for i in Insumo.objects.select_related("unidad_base").all()}
        alias_map = {_norm(a.nombre): a.insumo for a in InsumoAlias.objects.select_related("insumo").all()}

        wb = openpyxl.load_workbook(str(path), data_only=True)

        # ── INVENTARIO sheet ──────────────────────────────────────────────────
        ws_inv = wb["INVENTARIO"]
        updated_exist = 0
        created_exist = 0
        unmatched_inv = []

        for r in range(7, ws_inv.max_row + 1):
            nombre_raw = ws_inv.cell(r, 2).value
            if not nombre_raw or str(nombre_raw).strip().upper() in SKIP_NAMES:
                continue
            nombre = str(nombre_raw).strip()
            almacen_raw = str(ws_inv.cell(r, 5).value or "").strip()
            almacen = ALMACEN_MAP.get(almacen_raw, "ALMACEN_1")

            insumo = _match_insumo(nombre, nombre_map, alias_map)
            if insumo is None:
                unmatched_inv.append(nombre)
                continue

            stock_actual = _d(ws_inv.cell(r, 9).value)
            stock_min = _d(ws_inv.cell(r, 13).value)
            stock_max = _d(ws_inv.cell(r, 14).value)
            punto_reorden = _d(ws_inv.cell(r, 18).value)
            dias_llegada = int(_d(ws_inv.cell(r, 19).value))
            consumo_diario = _d(ws_inv.cell(r, 20).value)
            inv_prom = _d(ws_inv.cell(r, 16).value)

            if not dry_run:
                obj, created = ExistenciaInsumo.objects.get_or_create(
                    insumo=insumo,
                    defaults={
                        "almacen": almacen,
                        "stock_actual": stock_actual,
                        "stock_minimo": stock_min,
                        "stock_maximo": stock_max,
                        "punto_reorden": punto_reorden,
                        "dias_llegada_pedido": dias_llegada,
                        "consumo_diario_promedio": consumo_diario,
                        "inventario_promedio": inv_prom,
                    },
                )
                if not created:
                    obj.almacen = almacen
                    obj.stock_actual = stock_actual
                    obj.stock_minimo = stock_min
                    obj.stock_maximo = stock_max
                    obj.punto_reorden = punto_reorden
                    obj.dias_llegada_pedido = dias_llegada
                    obj.consumo_diario_promedio = consumo_diario
                    obj.inventario_promedio = inv_prom
                    obj.actualizado_en = timezone.now()
                    obj.save()
                    updated_exist += 1
                else:
                    created_exist += 1

        self.stdout.write(
            f"ExistenciaInsumo: {created_exist} creados, {updated_exist} actualizados"
        )
        if unmatched_inv:
            self.stdout.write(
                self.style.WARNING(
                    f"  Sin match ({len(unmatched_inv)}): {', '.join(unmatched_inv[:10])}"
                    + (f"... +{len(unmatched_inv)-10}" if len(unmatched_inv) > 10 else "")
                )
            )

        # ── ENTRADAS / SALIDAS sheets ─────────────────────────────────────────
        created_mov = 0
        skipped_dup = 0

        def _process_movements(ws, tipo: str) -> None:
            nonlocal created_mov, skipped_dup
            # Build almacen lookup from INVENTARIO sheet (nombre → almacen)
            nombre_almacen = {}
            for r2 in range(7, ws_inv.max_row + 1):
                n = ws_inv.cell(r2, 2).value
                a = str(ws_inv.cell(r2, 5).value or "").strip()
                if n and str(n).strip().upper() not in SKIP_NAMES:
                    nombre_almacen[str(n).strip()] = ALMACEN_MAP.get(a, "ALMACEN_1")

            for r in range(7, ws.max_row + 1):
                nombre_raw = ws.cell(r, 1).value
                if not nombre_raw or str(nombre_raw).strip().upper() in SKIP_NAMES:
                    continue
                nombre = str(nombre_raw).strip()
                insumo = _match_insumo(nombre, nombre_map, alias_map)
                if insumo is None:
                    continue
                almacen = nombre_almacen.get(nombre, "ALMACEN_1")

                for day in range(1, 32):
                    val = ws.cell(r, day + 1).value
                    if not val:
                        continue
                    cantidad = _d(val)
                    if cantidad <= 0:
                        continue
                    try:
                        mov_date = date(year, month, day)
                    except ValueError:
                        continue

                    # Idempotency hash
                    h_str = f"{tipo}:{insumo.id}:{mov_date.isoformat()}:{cantidad}:EXCEL_ALMACEN"
                    source_hash = hashlib.sha256(h_str.encode()).hexdigest()

                    if MovimientoInventario.objects.filter(source_hash=source_hash).exists():
                        skipped_dup += 1
                        continue

                    if not dry_run:
                        MovimientoInventario.objects.create(
                            fecha=datetime(year, month, day, 8, 0, tzinfo=dt_timezone.utc),
                            tipo=tipo,
                            insumo=insumo,
                            cantidad=cantidad,
                            almacen=almacen,
                            referencia=f"EXCEL_{year}{month:02d}",
                            notas="Importado desde Excel almacén",
                            source_hash=source_hash,
                        )
                    created_mov += 1

        if "ENTRADAS" in wb.sheetnames:
            _process_movements(wb["ENTRADAS"], MovimientoInventario.TIPO_ENTRADA)
        if "SALIDAS" in wb.sheetnames:
            _process_movements(wb["SALIDAS"], MovimientoInventario.TIPO_SALIDA)

        self.stdout.write(
            f"MovimientoInventario: {created_mov} creados, {skipped_dup} duplicados omitidos"
        )

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN — ningún cambio guardado"))
        else:
            self.stdout.write(self.style.SUCCESS("Importación completada"))
