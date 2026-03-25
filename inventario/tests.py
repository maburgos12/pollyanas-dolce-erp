from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
from openpyxl import load_workbook
from io import BytesIO

from core.models import AuditLog
from compras.models import SolicitudCompra
from inventario.models import AjusteInventario, AlmacenSyncRun, ExistenciaInsumo, MovimientoInventario
from maestros.models import CostoInsumo, Insumo, InsumoAlias, PointPendingMatch, Proveedor, UnidadMedida
from recetas.models import LineaReceta, Receta, VentaHistorica


class InventarioAliasesPendingTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_inv",
            email="admin_inv@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

    def test_pending_recent_runs_empty_state_message(self):
        response = self.client.get(reverse("inventario:aliases_catalog"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No hay sincronizaciones recientes para mostrar en este entorno.")
        self.assertContains(response, reverse("inventario:carga_almacen"))
        self.assertContains(response, reverse("inventario:sync_drive_now"))

    def test_pending_persisted_hide_and_restore_visibility(self):
        run = AlmacenSyncRun.objects.create(
            source=AlmacenSyncRun.SOURCE_DRIVE,
            status=AlmacenSyncRun.STATUS_OK,
            started_at=timezone.now(),
            matched=10,
            unmatched=2,
            pending_preview=[
                {
                    "source": "inventario",
                    "row": 8,
                    "nombre_origen": "Harina pastelera 25kg",
                    "nombre_normalizado": "harina pastelera 25kg",
                    "sugerencia": "Harina pastelera",
                    "score": 92.0,
                }
            ],
        )

        response = self.client.get(reverse("inventario:aliases_catalog"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["pending_visible_count"], 1)
        self.assertEqual(response.context["pending_source"], "persisted")

        self.client.post(
            reverse("inventario:aliases_catalog"),
            {"action": "clear_pending", "hide_run_id": str(run.id)},
        )
        response_hidden = self.client.get(reverse("inventario:aliases_catalog"))
        self.assertEqual(response_hidden.status_code, 200)
        self.assertEqual(response_hidden.context["pending_visible_count"], 0)
        self.assertEqual(response_hidden.context["hidden_run_id"], run.id)
        self.assertIsNotNone(response_hidden.context["hidden_pending_run"])

        self.client.post(
            reverse("inventario:aliases_catalog"),
            {"action": "reset_hidden_pending"},
        )
        response_restored = self.client.get(reverse("inventario:aliases_catalog"))
        self.assertEqual(response_restored.status_code, 200)
        self.assertEqual(response_restored.context["pending_visible_count"], 1)
        self.assertEqual(response_restored.context["pending_source"], "persisted")

    def test_load_pending_run_moves_preview_to_session(self):
        run = AlmacenSyncRun.objects.create(
            source=AlmacenSyncRun.SOURCE_SCHEDULED,
            status=AlmacenSyncRun.STATUS_OK,
            started_at=timezone.now(),
            matched=20,
            unmatched=1,
            pending_preview=[
                {
                    "source": "inventario",
                    "row": 4,
                    "nombre_origen": "Mantequilla barra",
                    "nombre_normalizado": "mantequilla barra",
                    "sugerencia": "Mantequilla",
                    "score": 88.0,
                }
            ],
        )

        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {"action": "load_pending_run", "run_id": str(run.id)},
        )
        self.assertEqual(response.status_code, 302)
        session_preview = self.client.session.get("inventario_pending_preview")
        self.assertIsInstance(session_preview, list)
        self.assertEqual(len(session_preview), 1)

        response_after = self.client.get(reverse("inventario:aliases_catalog"))
        self.assertEqual(response_after.status_code, 200)
        self.assertEqual(response_after.context["pending_source"], "session")
        self.assertEqual(response_after.context["pending_visible_count"], 1)

    def test_export_cross_pending_csv(self):
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="P-100",
            point_nombre="Mantequilla Barra",
            fuzzy_score=88.5,
            fuzzy_sugerencia="Mantequilla",
        )
        receta = Receta.objects.create(nombre="Receta Test Export", hash_contenido="hash-export-001")
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=None,
            insumo_texto="Mantequilla Barra",
            cantidad=1,
            unidad=None,
            unidad_texto="kg",
            costo_unitario_snapshot=0,
            match_status=LineaReceta.STATUS_REJECTED,
        )

        response = self.client.get(
            reverse("inventario:aliases_catalog"),
            {"export": "cross_pending_csv"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        body = response.content.decode("utf-8")
        self.assertIn("nombre_muestra", body)
        self.assertIn("Mantequilla Barra", body)

    def test_export_cross_pending_xlsx(self):
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="P-101",
            point_nombre="Mantequilla Barra",
            fuzzy_score=88.5,
            fuzzy_sugerencia="Mantequilla",
        )
        receta = Receta.objects.create(nombre="Receta Test Export XLSX", hash_contenido="hash-export-xlsx-001")
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=None,
            insumo_texto="Mantequilla Barra",
            cantidad=1,
            unidad=None,
            unidad_texto="kg",
            costo_unitario_snapshot=0,
            match_status=LineaReceta.STATUS_REJECTED,
        )

        response = self.client.get(
            reverse("inventario:aliases_catalog"),
            {"export": "cross_pending_xlsx"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        wb = load_workbook(BytesIO(response.content), data_only=True)
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, 10)]
        self.assertEqual(
            headers,
            [
                "nombre_muestra",
                "nombre_normalizado",
                "point_count",
                "almacen_count",
                "receta_count",
                "fuentes_activas",
                "total_count",
                "sugerencia",
                "score_max",
            ],
        )
        self.assertEqual(ws.cell(row=2, column=1).value, "Mantequilla Barra")

    def test_carga_almacen_shows_pending_homologation_focus(self):
        session = self.client.session
        session["inventario_pending_preview"] = [
            {
                "source": "inventario",
                "row": 12,
                "nombre_origen": "Harina pastelera 25kg",
                "nombre_normalizado": "harina pastelera 25kg",
                "sugerencia": "Harina Pastelera",
                "score": 94.0,
            },
            {
                "source": "entradas",
                "row": 7,
                "nombre_origen": "Harina pastelera 25kg",
                "nombre_normalizado": "harina pastelera 25kg",
                "sugerencia": "Harina Pastelera",
                "score": 91.0,
            },
            {
                "source": "inventario",
                "row": 14,
                "nombre_origen": "Azucar glass premium",
                "nombre_normalizado": "azucar glass premium",
                "sugerencia": "Azucar Glass",
                "score": 89.0,
            },
        ]
        session.save()

        response = self.client.get(reverse("inventario:carga_almacen"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Bloqueo de catálogo prioritario")
        self.assertIn("pending_focus", response.context)
        self.assertEqual(response.context["pending_focus"]["tone"], "warning")
        self.assertEqual(response.context["pending_focus"]["label"], "2 filas · entradas, inventario")
        self.assertEqual(len(response.context["pending_source_cards"]), 2)
        self.assertGreaterEqual(len(response.context["pending_focus_rows"]), 2)
        first_row = response.context["pending_focus_rows"][0]
        self.assertEqual(first_row["nombre_origen"], "Harina pastelera 25kg")
        self.assertContains(response, reverse("inventario:aliases_catalog"))
        self.assertContains(response, "Harina pastelera 25kg")

    def test_aliases_catalog_shows_pending_homologation_focus(self):
        session = self.client.session
        session["inventario_pending_preview"] = [
            {
                "source": "inventario",
                "row": 9,
                "nombre_origen": "Mermelada fresa premium",
                "nombre_normalizado": "mermelada fresa premium",
                "sugerencia": "Mermelada de Fresa",
                "score": 93.0,
            },
            {
                "source": "recetas",
                "row": 4,
                "nombre_origen": "Mermelada fresa premium",
                "nombre_normalizado": "mermelada fresa premium",
                "sugerencia": "Mermelada de Fresa",
                "score": 89.0,
            },
        ]
        session.save()

        response = self.client.get(reverse("inventario:aliases_catalog"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Bloqueo ERP prioritario")
        self.assertEqual(response.context["pending_focus"]["tone"], "warning")
        self.assertEqual(response.context["pending_focus"]["label"], "2 filas · inventario, recetas")
        self.assertEqual(len(response.context["pending_source_cards"]), 2)
        self.assertEqual(response.context["pending_focus_rows"][0]["nombre_origen"], "Mermelada fresa premium")
        self.assertContains(response, "Mermelada fresa premium")

    def test_alias_targets_uses_canonicalized_insumos(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor Canon", activo=True)
        unidad = UnidadMedida.objects.create(
            codigo="kg-can",
            nombre="Kilogramo Canon",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        canonical = Insumo.objects.create(
            nombre="Harina Canonica Inventario",
            categoria="Masa",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            activo=True,
            codigo_point="INV-001",
        )
        variant = Insumo.objects.create(
            nombre="Harina Canonica Inventario ",
            categoria="Masa",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=canonical,
            proveedor=proveedor,
            costo_unitario=Decimal("12"),
            source_hash="inv-canonical-cost",
        )
        CostoInsumo.objects.create(
            insumo=variant,
            proveedor=proveedor,
            costo_unitario=Decimal("11"),
            source_hash="inv-variant-cost",
        )

        response = self.client.get(reverse("inventario:aliases_catalog"))
        self.assertEqual(response.status_code, 200)
        targets = response.context["insumo_alias_targets"]
        ids = {item.id for item in targets}
        self.assertIn(canonical.id, ids)
        self.assertNotIn(variant.id, ids)
        selected = next(item for item in targets if item.id == canonical.id)
        self.assertEqual(selected.canonical_variant_count, 2)

    def test_movimientos_uses_canonicalized_insumos_and_aggregated_stock(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor Mov Canon", activo=True)
        unidad = UnidadMedida.objects.create(
            codigo="pz-can-mov",
            nombre="Pieza Mov Canon",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        canonical = Insumo.objects.create(
            nombre="Caja Canonica Inventario",
            categoria="Empaque",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            activo=True,
            codigo_point="INV-MOV-001",
        )
        variant = Insumo.objects.create(
            nombre="Caja Canonica Inventario ",
            categoria="Empaque",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            activo=True,
        )
        ExistenciaInsumo.objects.create(insumo=canonical, stock_actual=Decimal("4"))
        ExistenciaInsumo.objects.create(insumo=variant, stock_actual=Decimal("6"))

        response = self.client.get(reverse("inventario:movimientos"))
        self.assertEqual(response.status_code, 200)
        options = response.context["insumo_options"]
        ids = {item["id"] for item in options}
        self.assertIn(canonical.id, ids)
        self.assertNotIn(variant.id, ids)
        selected = next(item for item in options if item["id"] == canonical.id)
        self.assertEqual(selected["stock"], Decimal("10"))
        self.assertEqual(selected["canonical_variant_count"], 2)
        self.assertEqual(selected["enterprise_status"], "Lista para operar")
        self.assertEqual(selected["enterprise_missing"], [])
        self.assertFalse(selected["is_operational_blocker"])
        self.assertContains(response, "estado ERP")

    def test_movimientos_shows_master_focus_for_operational_blockers(self):
        unidad = UnidadMedida.objects.create(
            codigo="pz-mov-focus",
            nombre="Pieza Mov Focus",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        insumo = Insumo.objects.create(
            nombre="Caja Focus Movimientos",
            categoria="Empaque",
            unidad_base=unidad,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        ExistenciaInsumo.objects.create(
            insumo=insumo,
            stock_actual=Decimal("7"),
            stock_minimo=Decimal("2"),
            stock_maximo=Decimal("9"),
            punto_reorden=Decimal("3"),
        )

        response = self.client.get(reverse("inventario:movimientos"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["master_focus"]["class_label"], "Empaque")
        self.assertEqual(response.context["master_focus"]["missing_field"], "maestro")
        self.assertTrue(response.context["master_focus_rows"])
        self.assertTrue(response.context["master_blocker_class_cards"])
        focus_row = response.context["master_focus_rows"][0]
        self.assertIn(f"insumo_id={insumo.id}", focus_row["action_url"])
        self.assertEqual(focus_row["edit_url"], reverse("maestros:insumo_update", args=[insumo.id]))
        self.assertContains(response, "Editar artículo")
        self.assertContains(response, "Bloqueo maestro prioritario")

    def test_movimientos_shows_enterprise_row_summary_and_actions(self):
        unidad = UnidadMedida.objects.create(
            codigo="pz-mov-row",
            nombre="Pieza Mov Row",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        insumo = Insumo.objects.create(
            nombre="Empaque Row Movimiento",
            categoria="Empaque",
            unidad_base=unidad,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        MovimientoInventario.objects.create(
            fecha=timezone.now(),
            tipo=MovimientoInventario.TIPO_ENTRADA,
            insumo=insumo,
            cantidad=Decimal("5"),
            referencia="ROW-ERP",
        )

        response = self.client.get(reverse("inventario:movimientos"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Uso ERP")
        self.assertContains(response, "Maestro ERP")
        self.assertContains(response, "Empaque final")
        self.assertContains(response, "Incompleto")
        self.assertContains(response, "Falta: código Point")
        self.assertContains(response, reverse("maestros:insumo_update", args=[insumo.id]))
        self.assertContains(response, f"insumo_id={insumo.id}")
        self.assertContains(response, "Ruta troncal ERP")
        self.assertContains(response, "Dependencias upstream ERP")
        self.assertContains(response, "Ruta crítica ERP")
        self.assertContains(response, "Tramo ERP")
        self.assertContains(response, "Con bloqueo")
        self.assertContains(response, "Dependencia")
        self.assertContains(response, "Salud operativa")
        self.assertIn("enterprise_chain", response.context)
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("upstream_dependency_rows", response.context)
        self.assertIn("dependency_status", response.context["enterprise_chain"][0])
        self.assertIn("operational_health_cards", response.context)

    def test_movimientos_shows_canonical_consolidation_action_for_variant_rows(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor Movimiento Canon", activo=True)
        unidad = UnidadMedida.objects.create(
            codigo="pz-mov-can",
            nombre="Pieza Movimiento Canon",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        canonical = Insumo.objects.create(
            nombre="Caja Canonica Movimientos",
            categoria="Empaque",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
            codigo_point="MOV-CAN-001",
        )
        variant = Insumo.objects.create(
            nombre="CAJA CANONICA MOVIMIENTOS",
            categoria="Empaque",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        MovimientoInventario.objects.create(
            fecha=timezone.now(),
            tipo=MovimientoInventario.TIPO_ENTRADA,
            insumo=variant,
            cantidad=Decimal("3"),
            referencia="CANON-MOV",
        )

        response = self.client.get(reverse("inventario:movimientos"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Referencia maestra: hay variantes activas con impacto en movimientos.")
        self.assertContains(response, "Consolidar")
        self.assertContains(response, canonical.nombre)
        self.assertContains(response, f"q={canonical.nombre.replace(' ', '+')}")

    def test_movimiento_post_normaliza_variante_al_canonico(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor Mov Post Canon", activo=True)
        unidad = UnidadMedida.objects.create(
            codigo="pz-post-can",
            nombre="Pieza Post Canon",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        canonical = Insumo.objects.create(
            nombre="Caja Canonica Movimiento",
            categoria="Empaque",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            activo=True,
            codigo_point="INV-MOV-POST-001",
        )
        variant = Insumo.objects.create(
            nombre="CAJA CANONICA MOVIMIENTO",
            categoria="Empaque",
            unidad_base=unidad,
            activo=True,
        )

        response = self.client.post(
            reverse("inventario:movimientos"),
            {
                "insumo_id": str(variant.id),
                "tipo": MovimientoInventario.TIPO_ENTRADA,
                "cantidad": "4",
                "fecha": timezone.now().isoformat(),
                "referencia": "POST-CANON",
            },
        )
        self.assertEqual(response.status_code, 302)
        movimiento = MovimientoInventario.objects.order_by("-id").first()
        self.assertEqual(movimiento.insumo_id, canonical.id)

    def test_alias_resolution_usa_costo_canonico_del_grupo(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor Alias Canon", activo=True)
        unidad = UnidadMedida.objects.create(
            codigo="kg-alias-can",
            nombre="Kilogramo Alias Canon",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        canonical = Insumo.objects.create(
            nombre="Ganache Canonico",
            categoria="Relleno",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            activo=True,
            codigo_point="INV-ALIAS-001",
        )
        variant = Insumo.objects.create(
            nombre="GANACHE CANONICO",
            categoria="Relleno",
            unidad_base=unidad,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=variant,
            proveedor=proveedor,
            costo_unitario=Decimal("41.750000"),
            source_hash="hash-alias-canonical-cost",
        )
        receta = Receta.objects.create(nombre="Receta Alias Canon", hash_contenido="hash-alias-can-001")
        linea = LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=None,
            insumo_texto="Ganache Canonico",
            cantidad=Decimal("1"),
            unidad=None,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("0"),
            match_status=LineaReceta.STATUS_REJECTED,
        )

        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {
                "action": "create",
                "alias_name": "Ganache Canonico",
                "insumo_id": str(canonical.id),
            },
        )
        self.assertEqual(response.status_code, 302)

        linea.refresh_from_db()
        self.assertEqual(linea.insumo_id, canonical.id)
        self.assertEqual(linea.match_status, LineaReceta.STATUS_AUTO)
        self.assertEqual(linea.costo_unitario_snapshot, Decimal("41.750000"))

    def test_existencias_uses_canonicalized_insumo_selector(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor Exist Canon", activo=True)
        unidad = UnidadMedida.objects.create(
            codigo="lt-can-ex",
            nombre="Litro Canon Exist",
            tipo=UnidadMedida.TIPO_VOLUMEN,
            factor_to_base=Decimal("1000"),
        )
        canonical = Insumo.objects.create(
            nombre="Leche Canonica Inventario",
            categoria="Volumen",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            activo=True,
            codigo_point="INV-EX-001",
        )
        variant = Insumo.objects.create(
            nombre="Leche Canonica Inventario ",
            categoria="Volumen",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            activo=True,
        )

        response = self.client.get(reverse("inventario:existencias"))
        self.assertEqual(response.status_code, 200)
        insumos = response.context["insumos"]
        ids = {item.id for item in insumos}
        self.assertIn(canonical.id, ids)
        self.assertNotIn(variant.id, ids)

    def test_existencias_aggregates_duplicate_variants_into_one_canonical_row(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor Existencias Canon", activo=True)
        unidad = UnidadMedida.objects.create(
            codigo="pz-ex-can",
            nombre="Pieza Existencias Canon",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        canonical = Insumo.objects.create(
            nombre="Caja Canonica Existencias",
            categoria="Empaque",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            activo=True,
            codigo_point="INV-EX-CAN-001",
        )
        variant = Insumo.objects.create(
            nombre="CAJA CANONICA EXISTENCIAS",
            categoria="Empaque",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            activo=True,
        )
        ExistenciaInsumo.objects.create(
            insumo=canonical,
            stock_actual=Decimal("4"),
            stock_minimo=Decimal("2"),
            stock_maximo=Decimal("10"),
            punto_reorden=Decimal("3"),
            inventario_promedio=Decimal("5"),
            dias_llegada_pedido=2,
            consumo_diario_promedio=Decimal("1"),
        )
        ExistenciaInsumo.objects.create(
            insumo=variant,
            stock_actual=Decimal("6"),
            stock_minimo=Decimal("1"),
            stock_maximo=Decimal("7"),
            punto_reorden=Decimal("2"),
            inventario_promedio=Decimal("4"),
            dias_llegada_pedido=1,
            consumo_diario_promedio=Decimal("2"),
        )

        response = self.client.get(reverse("inventario:existencias"))
        self.assertEqual(response.status_code, 200)
        existencias = response.context["existencias"]
        filtered = [row for row in existencias if row.insumo.id == canonical.id]
        self.assertEqual(len(filtered), 1)
        row = filtered[0]
        self.assertEqual(row.stock_actual, Decimal("10"))
        self.assertEqual(row.stock_minimo, Decimal("2"))
        self.assertEqual(row.stock_maximo, Decimal("10"))
        self.assertEqual(row.punto_reorden, Decimal("3"))
        self.assertEqual(row.inventario_promedio, Decimal("5"))
        self.assertEqual(row.dias_llegada_pedido, 2)
        self.assertEqual(row.consumo_diario_promedio, Decimal("1"))
        self.assertEqual(row.canonical_variant_count, 2)

    def test_existencias_shows_master_focus_for_operational_blockers(self):
        unidad = UnidadMedida.objects.create(
            codigo="pz-ex-focus",
            nombre="Pieza Exist Focus",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        insumo = Insumo.objects.create(
            nombre="Caja Focus Inventario",
            categoria="Empaque",
            unidad_base=unidad,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        ExistenciaInsumo.objects.create(
            insumo=insumo,
            stock_actual=Decimal("8"),
            stock_minimo=Decimal("2"),
            stock_maximo=Decimal("10"),
            punto_reorden=Decimal("3"),
        )

        response = self.client.get(reverse("inventario:existencias"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["master_focus"]["class_label"], "Empaque")
        self.assertEqual(response.context["master_focus"]["missing_field"], "maestro")
        self.assertTrue(response.context["master_focus_rows"])
        self.assertTrue(response.context["master_blocker_class_cards"])
        focus_row = response.context["master_focus_rows"][0]
        self.assertIn(f"insumo_id={insumo.id}", focus_row["action_url"])
        self.assertEqual(focus_row["edit_url"], reverse("maestros:insumo_update", args=[insumo.id]))
        self.assertContains(response, "Editar artículo")
        self.assertContains(response, "Bloqueo maestro prioritario")
        self.assertContains(response, "Bloquea inventario")

    def test_existencias_shows_enterprise_row_summary_and_actions(self):
        unidad = UnidadMedida.objects.create(
            codigo="pz-ex-row",
            nombre="Pieza Exist Row",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        insumo = Insumo.objects.create(
            nombre="Empaque Row Inventario",
            categoria="Empaque",
            unidad_base=unidad,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        ExistenciaInsumo.objects.create(
            insumo=insumo,
            stock_actual=Decimal("3"),
            stock_minimo=Decimal("2"),
            punto_reorden=Decimal("2"),
        )

        response = self.client.get(reverse("inventario:existencias"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Uso ERP")
        self.assertContains(response, "Maestro ERP")
        self.assertContains(response, "Empaque final")
        self.assertContains(response, "Incompleto")
        self.assertContains(response, "Editar artículo")
        self.assertContains(response, reverse("maestros:insumo_update", args=[insumo.id]))
        self.assertContains(response, f"insumo_id={insumo.id}")
        self.assertContains(response, "Cadena documental ERP")
        self.assertContains(response, "Ruta troncal ERP")
        self.assertContains(response, "Dependencias upstream ERP")
        self.assertContains(response, "Dependencia")
        self.assertContains(response, "Resumen de seguimiento")
        self.assertContains(response, "Cierre por etapa documental")
        self.assertContains(response, "Control por frente")
        self.assertContains(response, "Entrega de inventario a downstream")
        self.assertContains(response, "Tramo ERP")
        self.assertContains(response, "Con bloqueo")
        self.assertContains(response, "Salud operativa")
        self.assertIn("enterprise_chain", response.context)
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertIn("upstream_dependency_rows", response.context)
        self.assertIn("dependency_status", response.context["enterprise_chain"][0])
        self.assertIn("document_stage_rows", response.context)
        self.assertIn("erp_governance_rows", response.context)
        self.assertIn("downstream_handoff_rows", response.context)
        self.assertIn("operational_health_cards", response.context)

    def test_existencias_and_alertas_show_sales_demand_signal(self):
        unidad = UnidadMedida.objects.create(
            codigo="kg-inv-demand",
            nombre="Kg Inv Demand",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        insumo = Insumo.objects.create(
            nombre="Harina Demanda Inventario",
            categoria="Masa",
            unidad_base=unidad,
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            activo=True,
        )
        receta = Receta.objects.create(
            nombre="Pastel Demanda Inventario",
            hash_contenido="hash-demand-inventario-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=insumo,
            insumo_texto=insumo.nombre,
            cantidad=Decimal("2"),
            unidad=unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("10"),
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100,
            match_status=LineaReceta.STATUS_AUTO,
        )
        ExistenciaInsumo.objects.create(
            insumo=insumo,
            stock_actual=Decimal("1"),
            stock_minimo=Decimal("2"),
            punto_reorden=Decimal("3"),
        )
        VentaHistorica.objects.create(
            receta=receta,
            fecha=timezone.localdate() - timedelta(days=2),
            cantidad=Decimal("85"),
            tickets=5,
            monto_total=Decimal("900"),
        )

        response_exist = self.client.get(reverse("inventario:existencias"))
        self.assertEqual(response_exist.status_code, 200)
        self.assertContains(response_exist, "Señal histórica de demanda")
        self.assertContains(response_exist, "Semáforo comercial")
        self.assertContains(response_exist, "Años observados")
        self.assertContains(response_exist, "Temporadas comparables")
        self.assertContains(response_exist, "Control de demanda comercial")
        self.assertContains(response_exist, "Artículos prioritarios por demanda")
        self.assertContains(response_exist, "Aseguramiento comercial prioritario")
        self.assertContains(response_exist, "Demanda crítica bloqueada por maestro")
        self.assertContains(response_exist, "Liberación operativa retenida")
        self.assertContains(response_exist, "Faltante maestro")
        self.assertContains(response_exist, "Pastel Demanda Inventario")
        self.assertIn("sales_demand_signal", response_exist.context)
        self.assertIn("years_observed", response_exist.context["sales_demand_signal"])
        self.assertIn("comparable_years", response_exist.context["sales_demand_signal"])
        self.assertIn("sales_demand_gate", response_exist.context)
        self.assertIn("commercial_priority_rows", response_exist.context)
        self.assertTrue(response_exist.context["commercial_priority_rows"])
        self.assertIn("critical_master_demand_rows", response_exist.context)
        self.assertTrue(response_exist.context["critical_master_demand_rows"])
        self.assertIn("supply_focus_rows", response_exist.context)
        self.assertTrue(response_exist.context["supply_focus_rows"])
        self.assertIn("daily_critical_close_focus", response_exist.context)
        self.assertIsNotNone(response_exist.context["daily_critical_close_focus"])
        self.assertContains(response_exist, "Cierre prioritario del día")

        response_alert = self.client.get(reverse("inventario:alertas"))
        self.assertEqual(response_alert.status_code, 200)
        self.assertContains(response_alert, "Señal histórica de demanda")
        self.assertContains(response_alert, "Semáforo comercial")
        self.assertContains(response_alert, "Años observados")
        self.assertContains(response_alert, "Temporadas comparables")
        self.assertContains(response_alert, "Control de demanda comercial")
        self.assertContains(response_alert, "Artículos prioritarios por demanda")
        self.assertContains(response_alert, "Aseguramiento comercial prioritario")
        self.assertContains(response_alert, "Demanda crítica bloqueada por maestro")
        self.assertContains(response_alert, "Liberación operativa retenida")
        self.assertContains(response_alert, "Faltante maestro")
        self.assertContains(response_alert, insumo.nombre)
        self.assertIn("sales_demand_signal", response_alert.context)
        self.assertIn("years_observed", response_alert.context["sales_demand_signal"])
        self.assertIn("comparable_years", response_alert.context["sales_demand_signal"])
        self.assertIn("sales_demand_gate", response_alert.context)
        self.assertIn("critical_master_demand_rows", response_alert.context)
        self.assertTrue(response_alert.context["critical_master_demand_rows"])
        self.assertIn("supply_focus_rows", response_alert.context)
        self.assertTrue(response_alert.context["supply_focus_rows"])
        self.assertIn("daily_critical_close_focus", response_alert.context)
        self.assertIsNotNone(response_alert.context["daily_critical_close_focus"])
        self.assertContains(response_alert, "Cierre prioritario del día")

    def test_existencias_shows_canonical_consolidation_action_when_variants_have_inventory_impact(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor Exist Canon Ops", activo=True)
        unidad = UnidadMedida.objects.create(
            codigo="pz-ex-can-ops",
            nombre="Pieza Exist Can Ops",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        canonical = Insumo.objects.create(
            nombre="Caja Canonica Operativa Inventario",
            categoria="Empaque",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
            codigo_point="INV-CAN-OPS-001",
        )
        variant = Insumo.objects.create(
            nombre="CAJA CANONICA OPERATIVA INVENTARIO",
            categoria="Empaque",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        ExistenciaInsumo.objects.create(
            insumo=canonical,
            stock_actual=Decimal("4"),
            stock_minimo=Decimal("2"),
            punto_reorden=Decimal("2"),
        )
        ExistenciaInsumo.objects.create(
            insumo=variant,
            stock_actual=Decimal("6"),
            stock_minimo=Decimal("2"),
            punto_reorden=Decimal("2"),
        )

        response = self.client.get(reverse("inventario:existencias"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Referencia maestra: existen variantes activas con impacto en inventario.")
        self.assertContains(response, "Consolidar")
        self.assertContains(response, "canonical_status=variantes")

    def test_existencias_can_focus_master_blocker_group(self):
        unidad = UnidadMedida.objects.create(
            codigo="pz-ex-focus-group",
            nombre="Pieza Exist Focus Group",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        empaque = Insumo.objects.create(
            nombre="Caja Focus Group",
            categoria="Empaque",
            unidad_base=unidad,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        mp = Insumo.objects.create(
            nombre="Azucar Focus Group",
            categoria="Azucar",
            unidad_base=unidad,
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            activo=True,
        )
        ExistenciaInsumo.objects.create(insumo=empaque, stock_actual=Decimal("2"), stock_minimo=Decimal("1"), punto_reorden=Decimal("1"))
        ExistenciaInsumo.objects.create(insumo=mp, stock_actual=Decimal("2"), stock_minimo=Decimal("1"), punto_reorden=Decimal("1"))

        response = self.client.get(reverse("inventario:existencias"), {"master_focus_key": "EMPAQUE:codigo_point"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_master_focus_key"], "EMPAQUE:codigo_point")
        self.assertTrue(response.context["master_focus_rows"])
        self.assertTrue(any(row["class_label"] == "Empaque" for row in response.context["master_focus_rows"]))
        self.assertContains(response, "Vista enfocada")

    def test_alertas_shows_master_focus_for_operational_blockers(self):
        unidad = UnidadMedida.objects.create(
            codigo="pz-alert-focus",
            nombre="Pieza Alert Focus",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        insumo = Insumo.objects.create(
            nombre="Caja Focus Alertas",
            categoria="Empaque",
            unidad_base=unidad,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        ExistenciaInsumo.objects.create(
            insumo=insumo,
            stock_actual=Decimal("0"),
            stock_minimo=Decimal("2"),
            stock_maximo=Decimal("10"),
            punto_reorden=Decimal("3"),
        )

        response = self.client.get(reverse("inventario:alertas"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["master_focus"]["class_label"], "Empaque")
        self.assertEqual(response.context["master_focus"]["missing_field"], "maestro")
        self.assertTrue(response.context["master_focus_rows"])
        self.assertTrue(response.context["master_blocker_class_cards"])
        focus_row = response.context["master_focus_rows"][0]
        self.assertIn(f"insumo_id={insumo.id}", focus_row["action_url"])
        self.assertEqual(focus_row["edit_url"], reverse("maestros:insumo_update", args=[insumo.id]))
        self.assertContains(response, "Editar artículo")
        self.assertContains(response, "Bloqueo maestro prioritario")
        self.assertContains(response, "Bloquea inventario")

    def test_alertas_shows_enterprise_row_summary_and_actions(self):
        unidad = UnidadMedida.objects.create(
            codigo="pz-alert-row",
            nombre="Pieza Alert Row",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        insumo = Insumo.objects.create(
            nombre="Empaque Row Alerta",
            categoria="Empaque",
            unidad_base=unidad,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        ExistenciaInsumo.objects.create(
            insumo=insumo,
            stock_actual=Decimal("0"),
            stock_minimo=Decimal("2"),
            punto_reorden=Decimal("1"),
        )

        response = self.client.get(reverse("inventario:alertas"), {"nivel": "all"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Uso ERP")
        self.assertContains(response, "Maestro ERP")
        self.assertContains(response, "Empaque final")
        self.assertContains(response, "Incompleto")
        self.assertContains(response, "Editar artículo")
        self.assertContains(response, reverse("maestros:insumo_update", args=[insumo.id]))
        self.assertContains(response, f"insumo_id={insumo.id}")
        self.assertContains(response, "Cadena documental ERP")
        self.assertContains(response, "Ruta troncal ERP")
        self.assertContains(response, "Dependencias upstream ERP")
        self.assertContains(response, "Prioridades de atención")
        self.assertContains(response, "Resumen de seguimiento")
        self.assertContains(response, "Dependencia")
        self.assertContains(response, "Cierre por etapa documental")
        self.assertContains(response, "Entrega de inventario a downstream")
        self.assertContains(response, "Tramo ERP")
        self.assertContains(response, "Con bloqueo")
        self.assertContains(response, "Salud operativa ERP")
        self.assertIn("enterprise_chain", response.context)
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertIn("upstream_dependency_rows", response.context)
        self.assertIn("dependency_status", response.context["enterprise_chain"][0])
        self.assertIn("document_stage_rows", response.context)
        self.assertIn("downstream_handoff_rows", response.context)
        self.assertIn("operational_health_cards", response.context)

    def test_alertas_preserves_level_when_master_focus_is_applied(self):
        unidad = UnidadMedida.objects.create(
            codigo="pz-alert-focus-group",
            nombre="Pieza Alert Focus Group",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        insumo = Insumo.objects.create(
            nombre="Caja Focus Alertas Group",
            categoria="Empaque",
            unidad_base=unidad,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        ExistenciaInsumo.objects.create(
            insumo=insumo,
            stock_actual=Decimal("0"),
            stock_minimo=Decimal("2"),
            stock_maximo=Decimal("5"),
            punto_reorden=Decimal("3"),
        )

        response = self.client.get(reverse("inventario:alertas"), {"nivel": "critico", "master_focus_key": "EMPAQUE:codigo_point"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_master_focus_key"], "EMPAQUE:codigo_point")
        self.assertContains(response, "Vista enfocada")
        self.assertContains(response, "?nivel=critico")

    def test_ajustes_shows_master_focus_for_operational_blockers(self):
        unidad = UnidadMedida.objects.create(
            codigo="pz-aj-focus",
            nombre="Pieza Ajuste Focus",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        insumo = Insumo.objects.create(
            nombre="Caja Focus Ajustes",
            categoria="Empaque",
            unidad_base=unidad,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        ExistenciaInsumo.objects.create(
            insumo=insumo,
            stock_actual=Decimal("5"),
            stock_minimo=Decimal("2"),
            stock_maximo=Decimal("8"),
            punto_reorden=Decimal("3"),
        )

        response = self.client.get(reverse("inventario:ajustes"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["master_focus"]["class_label"], "Empaque")
        self.assertEqual(response.context["master_focus"]["missing_field"], "maestro")
        self.assertTrue(response.context["master_focus_rows"])
        self.assertTrue(response.context["master_blocker_class_cards"])
        focus_row = response.context["master_focus_rows"][0]
        self.assertIn(f"insumo_id={insumo.id}", focus_row["action_url"])
        self.assertEqual(focus_row["edit_url"], reverse("maestros:insumo_update", args=[insumo.id]))
        self.assertContains(response, "Editar artículo")
        self.assertContains(response, "Bloqueo maestro prioritario")

    def test_ajustes_shows_enterprise_row_summary_and_actions(self):
        unidad = UnidadMedida.objects.create(
            codigo="pz-aj-row",
            nombre="Pieza Aj Row",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        insumo = Insumo.objects.create(
            nombre="Empaque Row Ajuste",
            categoria="Empaque",
            unidad_base=unidad,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        ajuste = AjusteInventario.objects.create(
            insumo=insumo,
            cantidad_sistema=Decimal("5"),
            cantidad_fisica=Decimal("4"),
            motivo="ROW-ERP",
            estatus=AjusteInventario.STATUS_PENDIENTE,
            solicitado_por=self.user,
        )

        response = self.client.get(reverse("inventario:ajustes"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Uso ERP")
        self.assertContains(response, "Maestro ERP")
        self.assertContains(response, "Empaque final")
        self.assertContains(response, "Incompleto")
        self.assertContains(response, "Falta: código Point")
        self.assertContains(response, reverse("maestros:insumo_update", args=[insumo.id]))
        self.assertContains(response, f"insumo_id={ajuste.insumo_id}")
        self.assertContains(response, "Cadena documental ERP")
        self.assertContains(response, "Ruta troncal ERP")
        self.assertContains(response, "Dependencias upstream ERP")
        self.assertContains(response, "Ruta crítica ERP")
        self.assertContains(response, "Radar ejecutivo ERP")
        self.assertContains(response, "Dependencia")
        self.assertContains(response, "Cierre por etapa documental")
        self.assertContains(response, "Entrega de inventario a downstream")
        self.assertContains(response, "Tramo ERP")
        self.assertContains(response, "Con bloqueo")
        self.assertContains(response, "Salud operativa ERP")
        self.assertIn("enterprise_chain", response.context)
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertIn("upstream_dependency_rows", response.context)
        self.assertIn("dependency_status", response.context["enterprise_chain"][0])
        self.assertIn("document_stage_rows", response.context)
        self.assertIn("downstream_handoff_rows", response.context)
        self.assertIn("operational_health_cards", response.context)

    def test_ajustes_shows_canonical_consolidation_action_for_variant_rows(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor Ajuste Canon", activo=True)
        unidad = UnidadMedida.objects.create(
            codigo="pz-aj-can",
            nombre="Pieza Ajuste Canon",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        canonical = Insumo.objects.create(
            nombre="Caja Canonica Ajustes",
            categoria="Empaque",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
            codigo_point="AJ-CAN-001",
        )
        variant = Insumo.objects.create(
            nombre="CAJA CANONICA AJUSTES",
            categoria="Empaque",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        ajuste = AjusteInventario.objects.create(
            insumo=variant,
            cantidad_sistema=Decimal("5"),
            cantidad_fisica=Decimal("4"),
            motivo="CANON",
            estatus=AjusteInventario.STATUS_PENDIENTE,
            solicitado_por=self.user,
        )

        response = self.client.get(reverse("inventario:ajustes"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Consolidar")
        self.assertContains(response, "Consolidar")
        self.assertContains(response, canonical.nombre)
        self.assertContains(response, f"q={canonical.nombre.replace(' ', '+')}")

    def test_export_cross_pending_csv_with_filters(self):
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="P-200",
            point_nombre="Mantequilla Barra",
            fuzzy_score=95.0,
            fuzzy_sugerencia="Mantequilla",
        )
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="P-201",
            point_nombre="Azucar Morena",
            fuzzy_score=72.0,
            fuzzy_sugerencia="Azucar",
        )
        receta = Receta.objects.create(nombre="Receta Test Filtros", hash_contenido="hash-export-002")
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=None,
            insumo_texto="Mantequilla Barra",
            cantidad=1,
            unidad=None,
            unidad_texto="kg",
            costo_unitario_snapshot=0,
            match_status=LineaReceta.STATUS_REJECTED,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=2,
            insumo=None,
            insumo_texto="Azucar Morena",
            cantidad=1,
            unidad=None,
            unidad_texto="kg",
            costo_unitario_snapshot=0,
            match_status=LineaReceta.STATUS_REJECTED,
        )

        response = self.client.get(
            reverse("inventario:aliases_catalog"),
            {
                "export": "cross_pending_csv",
                "cross_q": "mantequilla",
                "cross_min_sources": "2",
                "cross_score_min": "90",
                "cross_only_suggested": "1",
            },
        )
        self.assertEqual(response.status_code, 200)
        body = response.content.decode("utf-8")
        self.assertIn("Mantequilla Barra", body)
        self.assertNotIn("Azucar Morena", body)

    def test_export_cross_pending_csv_with_point_tipo_filter(self):
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="P-TIPO-INS-001",
            point_nombre="Insumo only row",
            fuzzy_score=80.0,
            fuzzy_sugerencia="Harina",
        )
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_PRODUCTO,
            point_codigo="P-TIPO-PROD-001",
            point_nombre="Producto only row",
            fuzzy_score=86.0,
            fuzzy_sugerencia="Receta sugerida",
        )

        response = self.client.get(
            reverse("inventario:aliases_catalog"),
            {
                "export": "cross_pending_csv",
                "cross_point_tipo": "PRODUCTO",
                "cross_min_sources": "1",
            },
        )
        self.assertEqual(response.status_code, 200)
        body = response.content.decode("utf-8")
        self.assertIn("Producto only row", body)
        self.assertNotIn("Insumo only row", body)

    def test_export_cross_pending_csv_with_source_filter(self):
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="P-SRC-POINT-001",
            point_nombre="Solo Point Export",
            fuzzy_score=91.0,
            fuzzy_sugerencia="Harina",
        )
        receta = Receta.objects.create(nombre="Receta Solo Source Export", hash_contenido="hash-export-source-001")
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=None,
            insumo_texto="Solo Receta Export",
            cantidad=1,
            unidad=None,
            unidad_texto="kg",
            costo_unitario_snapshot=0,
            match_status=LineaReceta.STATUS_REJECTED,
        )

        response = self.client.get(
            reverse("inventario:aliases_catalog"),
            {
                "export": "cross_pending_csv",
                "cross_source": "POINT",
                "cross_point_tipo": "INSUMO",
                "cross_min_sources": "1",
            },
        )
        self.assertEqual(response.status_code, 200)
        body = response.content.decode("utf-8")
        self.assertIn("Solo Point Export", body)
        self.assertNotIn("Solo Receta Export", body)

    def test_auto_apply_suggestions_creates_alias_and_cleans_pending(self):
        unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA)
        insumo = Insumo.objects.create(nombre="Harina Pastelera", unidad_base=unidad)
        run = AlmacenSyncRun.objects.create(
            source=AlmacenSyncRun.SOURCE_DRIVE,
            status=AlmacenSyncRun.STATUS_OK,
            started_at=timezone.now(),
            matched=11,
            unmatched=1,
            pending_preview=[
                {
                    "source": "inventario",
                    "row": 9,
                    "nombre_origen": "Harina pastelera 25kg",
                    "nombre_normalizado": "harina pastelera 25kg",
                    "sugerencia": "Harina Pastelera",
                    "score": 95.0,
                }
            ],
        )

        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {
                "action": "auto_apply_suggestions",
                "auto_min_score": "90",
                "auto_min_sources": "1",
                "auto_max_rows": "50",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            InsumoAlias.objects.filter(nombre_normalizado="harina pastelera 25kg", insumo=insumo).exists()
        )

        run.refresh_from_db()
        self.assertEqual(run.pending_preview, [])

    def test_auto_apply_suggestions_min_sources_gate(self):
        unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA)
        insumo = Insumo.objects.create(nombre="Azucar Glass", unidad_base=unidad)
        AlmacenSyncRun.objects.create(
            source=AlmacenSyncRun.SOURCE_DRIVE,
            status=AlmacenSyncRun.STATUS_OK,
            started_at=timezone.now(),
            matched=11,
            unmatched=1,
            pending_preview=[
                {
                    "source": "inventario",
                    "row": 5,
                    "nombre_origen": "Azucar glass fina",
                    "nombre_normalizado": "azucar glass fina",
                    "sugerencia": "Azucar Glass",
                    "score": 96.0,
                }
            ],
        )

        # Con 2 fuentes mínimas no debe crear alias (solo existe en almacén).
        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {
                "action": "auto_apply_suggestions",
                "auto_min_score": "90",
                "auto_min_sources": "2",
                "auto_max_rows": "50",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            InsumoAlias.objects.filter(nombre_normalizado="azucar glass fina", insumo=insumo).exists()
        )

        # Agregamos pendiente en Point para activar 2 fuentes y sí debe crear alias.
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="P-XYZ",
            point_nombre="Azucar glass fina",
            fuzzy_score=96.0,
            fuzzy_sugerencia="Azucar Glass",
        )
        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {
                "action": "auto_apply_suggestions",
                "auto_min_score": "90",
                "auto_min_sources": "2",
                "auto_max_rows": "50",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            InsumoAlias.objects.filter(nombre_normalizado="azucar glass fina", insumo=insumo).exists()
        )

    def test_auto_apply_suggestions_respects_cross_filters(self):
        unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA)
        insumo_mantequilla = Insumo.objects.create(nombre="Mantequilla", unidad_base=unidad)
        insumo_azucar = Insumo.objects.create(nombre="Azucar Morena", unidad_base=unidad)

        AlmacenSyncRun.objects.create(
            source=AlmacenSyncRun.SOURCE_DRIVE,
            status=AlmacenSyncRun.STATUS_OK,
            started_at=timezone.now(),
            matched=20,
            unmatched=2,
            pending_preview=[
                {
                    "source": "inventario",
                    "row": 7,
                    "nombre_origen": "Mantequilla barra",
                    "nombre_normalizado": "mantequilla barra",
                    "sugerencia": "Mantequilla",
                    "score": 95.0,
                },
                {
                    "source": "inventario",
                    "row": 8,
                    "nombre_origen": "Azucar morena premium",
                    "nombre_normalizado": "azucar morena premium",
                    "sugerencia": "Azucar Morena",
                    "score": 96.0,
                },
            ],
        )
        # Damos 2 fuentes activas a ambas filas para evitar filtros por fuente.
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PM-001",
            point_nombre="Mantequilla barra",
            fuzzy_score=95.0,
            fuzzy_sugerencia="Mantequilla",
        )
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PM-002",
            point_nombre="Azucar morena premium",
            fuzzy_score=96.0,
            fuzzy_sugerencia="Azucar Morena",
        )

        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {
                "action": "auto_apply_suggestions",
                "auto_min_score": "90",
                "auto_min_sources": "1",
                "auto_max_rows": "50",
                "cross_q": "mantequilla",
                "cross_min_sources": "1",
                "cross_score_min": "0",
                "cross_only_suggested": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            InsumoAlias.objects.filter(nombre_normalizado="mantequilla barra", insumo=insumo_mantequilla).exists()
        )
        self.assertFalse(
            InsumoAlias.objects.filter(nombre_normalizado="azucar morena premium", insumo=insumo_azucar).exists()
        )

    def test_auto_apply_suggestions_respects_cross_source(self):
        unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA)
        insumo_point = Insumo.objects.create(nombre="Insumo Point", unidad_base=unidad)
        insumo_almacen = Insumo.objects.create(nombre="Insumo Almacen", unidad_base=unidad)

        AlmacenSyncRun.objects.create(
            source=AlmacenSyncRun.SOURCE_DRIVE,
            status=AlmacenSyncRun.STATUS_OK,
            started_at=timezone.now(),
            matched=20,
            unmatched=1,
            pending_preview=[
                {
                    "source": "inventario",
                    "row": 4,
                    "nombre_origen": "Solo almacen auto",
                    "nombre_normalizado": "solo almacen auto",
                    "sugerencia": "Insumo Almacen",
                    "score": 98.0,
                }
            ],
        )
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-AUTO-SRC-01",
            point_nombre="Solo point auto",
            fuzzy_score=98.0,
            fuzzy_sugerencia="Insumo Point",
        )

        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {
                "action": "auto_apply_suggestions",
                "auto_min_score": "90",
                "auto_min_sources": "1",
                "auto_max_rows": "50",
                "cross_min_sources": "1",
                "cross_score_min": "0",
                "cross_point_tipo": "INSUMO",
                "cross_source": "POINT",
                "cross_limit": "50",
                "cross_offset": "0",
                "cross_sort_by": "score_max",
                "cross_sort_dir": "desc",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(InsumoAlias.objects.filter(nombre_normalizado="solo point auto", insumo=insumo_point).exists())
        self.assertFalse(
            InsumoAlias.objects.filter(nombre_normalizado="solo almacen auto", insumo=insumo_almacen).exists()
        )

    def test_auto_apply_suggestions_logs_filters_payload(self):
        unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA)
        insumo = Insumo.objects.create(nombre="Mantequilla", unidad_base=unidad)
        AlmacenSyncRun.objects.create(
            source=AlmacenSyncRun.SOURCE_DRIVE,
            status=AlmacenSyncRun.STATUS_OK,
            started_at=timezone.now(),
            matched=11,
            unmatched=1,
            pending_preview=[
                {
                    "source": "inventario",
                    "row": 2,
                    "nombre_origen": "Mantequilla test log",
                    "nombre_normalizado": "mantequilla test log",
                    "sugerencia": "Mantequilla",
                    "score": 97.0,
                }
            ],
        )

        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {
                "action": "auto_apply_suggestions",
                "auto_min_score": "90",
                "auto_min_sources": "1",
                "auto_max_rows": "50",
                "cross_q": "mantequilla",
                "cross_source": "TODOS",
                "cross_point_tipo": "INSUMO",
                "cross_min_sources": "1",
                "cross_score_min": "0",
                "cross_limit": "50",
                "cross_offset": "0",
                "cross_sort_by": "score_max",
                "cross_sort_dir": "desc",
            },
        )
        self.assertEqual(response.status_code, 302)
        log = AuditLog.objects.filter(action="AUTO_APPLY_SUGGESTIONS", model="inventario.InsumoAlias").first()
        self.assertIsNotNone(log)
        payload = log.payload
        self.assertEqual(payload["filters"]["cross_source"], "TODOS")
        self.assertEqual(payload["filters"]["cross_point_tipo"], "INSUMO")
        self.assertEqual(payload["filters"]["cross_sort_by"], "score_max")
        self.assertEqual(payload["filters"]["cross_sort_dir"], "desc")
        self.assertEqual(payload["filters"]["cross_limit"], 50)
        self.assertEqual(payload["filters"]["cross_offset"], 0)
        self.assertGreaterEqual(int(payload["summary"]["processed"]), 1)

    def test_auto_apply_suggestions_redirect_keeps_cross_filters(self):
        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {
                "action": "auto_apply_suggestions",
                "auto_min_score": "90",
                "auto_min_sources": "2",
                "auto_max_rows": "50",
                "next_q": "harina",
                "cross_q": "mantequilla",
                "cross_source": "POINT",
                "cross_min_sources": "2",
                "cross_score_min": "90",
                "cross_point_tipo": "PRODUCTO",
                "cross_limit": "50",
                "cross_offset": "100",
                "cross_sort_by": "score_max",
                "cross_sort_dir": "asc",
                "cross_only_suggested": "1",
            },
        )
        self.assertEqual(response.status_code, 302)
        location = response["Location"]
        self.assertIn("q=harina", location)
        self.assertIn("cross_q=mantequilla", location)
        self.assertIn("cross_source=POINT", location)
        self.assertIn("cross_min_sources=2", location)
        self.assertIn("cross_score_min=90.0", location)
        self.assertIn("cross_point_tipo=PRODUCTO", location)
        self.assertIn("cross_limit=50", location)
        self.assertIn("cross_offset=100", location)
        self.assertIn("cross_sort_by=score_max", location)
        self.assertIn("cross_sort_dir=asc", location)
        self.assertIn("cross_only_suggested=1", location)

    def test_cross_unified_view_pagination_and_sort(self):
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-CROSS-PAG-01",
            point_nombre="C Name",
            fuzzy_score=85.0,
            fuzzy_sugerencia="Sugerencia C",
        )
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-CROSS-PAG-02",
            point_nombre="A Name",
            fuzzy_score=85.0,
            fuzzy_sugerencia="Sugerencia A",
        )
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-CROSS-PAG-03",
            point_nombre="B Name",
            fuzzy_score=85.0,
            fuzzy_sugerencia="Sugerencia B",
        )

        response = self.client.get(
            reverse("inventario:aliases_catalog"),
            {
                "cross_point_tipo": "INSUMO",
                "cross_min_sources": "1",
                "cross_limit": "1",
                "cross_offset": "1",
                "cross_sort_by": "nombre_muestra",
                "cross_sort_dir": "asc",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["cross_limit"], 1)
        self.assertEqual(response.context["cross_offset"], 1)
        self.assertEqual(response.context["cross_sort_by"], "nombre_muestra")
        self.assertEqual(response.context["cross_sort_dir"], "asc")
        self.assertEqual(response.context["cross_returned_count"], 1)
        self.assertTrue(response.context["cross_has_prev"])
        self.assertTrue(response.context["cross_has_next"])
        rows = response.context["cross_unified_rows"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["nombre_muestra"], "B Name")

    def test_cross_unified_view_source_filter(self):
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-CROSS-SRC-01",
            point_nombre="Solo Point",
            fuzzy_score=90.0,
            fuzzy_sugerencia="Sugerencia Point",
        )
        receta = Receta.objects.create(nombre="Receta Solo Source", hash_contenido="hash-cross-source-001")
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=None,
            insumo_texto="Solo Receta",
            cantidad=1,
            unidad=None,
            unidad_texto="kg",
            costo_unitario_snapshot=0,
            match_status=LineaReceta.STATUS_REJECTED,
        )

        response_point = self.client.get(
            reverse("inventario:aliases_catalog"),
            {
                "cross_source": "POINT",
                "cross_point_tipo": "INSUMO",
                "cross_min_sources": "1",
                "cross_limit": "50",
            },
        )
        self.assertEqual(response_point.status_code, 200)
        rows_point = response_point.context["cross_unified_rows"]
        self.assertGreaterEqual(len(rows_point), 1)
        self.assertTrue(all(int(row.get("point_count") or 0) > 0 for row in rows_point))
        point_stats = response_point.context["cross_source_stats"]
        self.assertEqual(point_stats["point_rows"], len(rows_point))
        self.assertEqual(point_stats["almacen_rows"], 0)
        self.assertEqual(point_stats["receta_rows"], 0)

        response_recetas = self.client.get(
            reverse("inventario:aliases_catalog"),
            {
                "cross_source": "RECETAS",
                "cross_point_tipo": "INSUMO",
                "cross_min_sources": "1",
                "cross_limit": "50",
            },
        )
        self.assertEqual(response_recetas.status_code, 200)
        rows_recetas = response_recetas.context["cross_unified_rows"]
        self.assertGreaterEqual(len(rows_recetas), 1)
        self.assertTrue(all(int(row.get("receta_count") or 0) > 0 for row in rows_recetas))
        receta_stats = response_recetas.context["cross_source_stats"]
        self.assertEqual(receta_stats["receta_rows"], len(rows_recetas))
        self.assertEqual(receta_stats["point_rows"], 0)
        self.assertEqual(receta_stats["almacen_rows"], 0)
        self.assertIn("cross_source=POINT", response_point.context["cross_query_source_point"])
        self.assertIn("cross_source=ALMACEN", response_point.context["cross_query_source_almacen"])
        self.assertIn("cross_source=RECETAS", response_point.context["cross_query_source_recetas"])
        self.assertIn("cross_source=TODOS", response_point.context["cross_query_source_todos"])

    def test_bulk_reassign_resolves_and_cleans_pending(self):
        unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA)
        insumo_a = Insumo.objects.create(nombre="Harina A", unidad_base=unidad)
        insumo_b = Insumo.objects.create(nombre="Harina B", unidad_base=unidad)
        alias = InsumoAlias.objects.create(nombre="Harina pastelera 25kg", insumo=insumo_a)
        run = AlmacenSyncRun.objects.create(
            source=AlmacenSyncRun.SOURCE_DRIVE,
            status=AlmacenSyncRun.STATUS_OK,
            started_at=timezone.now(),
            matched=7,
            unmatched=1,
            pending_preview=[
                {
                    "source": "inventario",
                    "row": 14,
                    "nombre_origen": "Harina pastelera 25kg",
                    "nombre_normalizado": "harina pastelera 25kg",
                    "sugerencia": "Harina B",
                    "score": 92.0,
                }
            ],
        )

        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {
                "action": "bulk_reassign",
                "insumo_id": str(insumo_b.id),
                "alias_ids": [str(alias.id)],
            },
        )
        self.assertEqual(response.status_code, 302)

        alias.refresh_from_db()
        self.assertEqual(alias.insumo_id, insumo_b.id)
        run.refresh_from_db()
        self.assertEqual(run.pending_preview, [])

    def test_bulk_import_aliases_creates_and_cleans_pending(self):
        unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA)
        insumo = Insumo.objects.create(nombre="Harina Pastelera", unidad_base=unidad)
        run = AlmacenSyncRun.objects.create(
            source=AlmacenSyncRun.SOURCE_DRIVE,
            status=AlmacenSyncRun.STATUS_OK,
            started_at=timezone.now(),
            matched=5,
            unmatched=1,
            pending_preview=[
                {
                    "source": "inventario",
                    "row": 6,
                    "nombre_origen": "Harina pastelera 25kg",
                    "nombre_normalizado": "harina pastelera 25kg",
                    "sugerencia": "Harina Pastelera",
                    "score": 95.0,
                }
            ],
        )
        payload = "alias,insumo\nHarina pastelera 25kg,Harina Pastelera\n"
        archivo = SimpleUploadedFile("aliases.csv", payload.encode("utf-8"), content_type="text/csv")

        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {
                "action": "import_bulk",
                "score_min": "90",
                "archivo_aliases": archivo,
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            InsumoAlias.objects.filter(nombre_normalizado="harina pastelera 25kg", insumo=insumo).exists()
        )
        run.refresh_from_db()
        self.assertEqual(run.pending_preview, [])

    def test_bulk_import_aliases_stores_unresolved_preview(self):
        payload = "alias,insumo\nFresa natural premium,No Existe En Catalogo\n"
        archivo = SimpleUploadedFile("aliases.csv", payload.encode("utf-8"), content_type="text/csv")

        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {
                "action": "import_bulk",
                "score_min": "95",
                "archivo_aliases": archivo,
            },
        )
        self.assertEqual(response.status_code, 302)

        preview = self.client.session.get("inventario_alias_import_preview")
        stats = self.client.session.get("inventario_alias_import_stats")
        self.assertIsInstance(preview, list)
        self.assertEqual(len(preview), 1)
        self.assertEqual(preview[0]["alias"], "Fresa natural premium")
        self.assertEqual(stats["unresolved"], 1)

    def test_alias_import_preview_export_csv_and_xlsx(self):
        session = self.client.session
        session["inventario_alias_import_preview"] = [
            {
                "row": 2,
                "alias": "Fresa natural premium",
                "insumo_archivo": "No Existe En Catalogo",
                "sugerencia": "Fresa Fresca",
                "score": 88.5,
                "method": "FUZZY",
                "motivo": "Insumo no resuelto (score<90.0).",
            }
        ]
        session.save()

        response_csv = self.client.get(reverse("inventario:aliases_catalog"), {"export": "alias_import_preview_csv"})
        self.assertEqual(response_csv.status_code, 200)
        self.assertIn("text/csv", response_csv["Content-Type"])
        body_csv = response_csv.content.decode("utf-8")
        self.assertIn("row,alias,insumo_archivo,sugerencia,score,method,motivo", body_csv)
        self.assertIn("Fresa natural premium", body_csv)

        response_xlsx = self.client.get(reverse("inventario:aliases_catalog"), {"export": "alias_import_preview_xlsx"})
        self.assertEqual(response_xlsx.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response_xlsx["Content-Type"],
        )
        wb = load_workbook(BytesIO(response_xlsx.content), data_only=True)
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, 8)]
        self.assertEqual(headers, ["row", "alias", "insumo_archivo", "sugerencia", "score", "method", "motivo"])
        self.assertEqual(ws.cell(row=2, column=2).value, "Fresa natural premium")

    def test_alias_template_export_csv_and_xlsx(self):
        response_csv = self.client.get(reverse("inventario:aliases_catalog"), {"export": "alias_template_csv"})
        self.assertEqual(response_csv.status_code, 200)
        self.assertIn("text/csv", response_csv["Content-Type"])
        self.assertIn("alias,insumo", response_csv.content.decode("utf-8"))

        response_xlsx = self.client.get(reverse("inventario:aliases_catalog"), {"export": "alias_template_xlsx"})
        self.assertEqual(response_xlsx.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response_xlsx["Content-Type"],
        )

    def test_apply_suggestion_creates_alias_and_cleans_pending(self):
        unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA)
        insumo = Insumo.objects.create(nombre="Mantequilla", unidad_base=unidad)
        run = AlmacenSyncRun.objects.create(
            source=AlmacenSyncRun.SOURCE_DRIVE,
            status=AlmacenSyncRun.STATUS_OK,
            started_at=timezone.now(),
            matched=12,
            unmatched=1,
            pending_preview=[
                {
                    "source": "inventario",
                    "row": 3,
                    "nombre_origen": "Mantequilla barra",
                    "nombre_normalizado": "mantequilla barra",
                    "sugerencia": "Mantequilla",
                    "score": 95.0,
                }
            ],
        )

        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {
                "action": "apply_suggestion",
                "alias_name": "Mantequilla barra",
                "suggestion": "Mantequilla",
                "score_min": "90",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            InsumoAlias.objects.filter(nombre_normalizado="mantequilla barra", insumo=insumo).exists()
        )
        run.refresh_from_db()
        self.assertEqual(run.pending_preview, [])

    def test_aliases_catalog_export_csv_and_xlsx(self):
        unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA)
        insumo = Insumo.objects.create(nombre="Harina", unidad_base=unidad)
        InsumoAlias.objects.create(nombre="Harina pastelera 25kg", insumo=insumo)

        response_csv = self.client.get(reverse("inventario:aliases_catalog"), {"export": "aliases_csv"})
        self.assertEqual(response_csv.status_code, 200)
        self.assertIn("text/csv", response_csv["Content-Type"])
        body = response_csv.content.decode("utf-8")
        self.assertIn("alias,normalizado,insumo_oficial", body)
        self.assertIn("Harina pastelera 25kg", body)

        response_xlsx = self.client.get(reverse("inventario:aliases_catalog"), {"export": "aliases_xlsx"})
        self.assertEqual(response_xlsx.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response_xlsx["Content-Type"],
        )
        wb = load_workbook(BytesIO(response_xlsx.content), data_only=True)
        ws = wb.active
        headers = [ws.cell(row=1, column=1).value, ws.cell(row=1, column=2).value, ws.cell(row=1, column=3).value]
        self.assertEqual(headers, ["alias", "normalizado", "insumo_oficial"])

    def test_aliases_catalog_renders_master_data_blocks(self):
        unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA)
        Insumo.objects.create(nombre="Azucar Normal", unidad_base=unidad)
        response = self.client.get(reverse("inventario:aliases_catalog"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Centro de mando ERP")
        self.assertContains(response, "Workflow ERP de referencias")
        self.assertContains(response, "Mesa de gobierno ERP")
        self.assertContains(response, "Gobierno operativo de referencias")
        self.assertContains(response, "Siguiente paso ERP")
        self.assertIn("erp_command_center", response.context)
        self.assertIn("erp_governance_rows", response.context)
        self.assertIn("master_normalize", response.context)
        self.assertIn("master_duplicates", response.context)
        self.assertIn("totales", response.context["master_normalize"])
        self.assertIn("totales", response.context["master_duplicates"])

    def test_master_normalize_apply_updates_nombre_normalizado(self):
        unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA)
        insumo = Insumo.objects.create(
            nombre="Azúcar Glass",
            nombre_normalizado="valor-invalido",
            unidad_base=unidad,
        )
        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {
                "action": "master_normalize",
                "master_scope": "insumos",
                "master_limit": "100",
                "master_offset": "0",
                "master_mode": "apply",
            },
        )
        self.assertEqual(response.status_code, 302)
        insumo.refresh_from_db()
        self.assertEqual(insumo.nombre_normalizado, "azucar glass")

    def test_master_duplicates_export_csv_and_xlsx_from_aliases_catalog(self):
        unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA)
        Insumo.objects.create(nombre="Mantequilla Barra", unidad_base=unidad)
        Insumo.objects.create(nombre="MANTEQUILLA  BARRA", unidad_base=unidad)

        response_csv = self.client.get(
            reverse("inventario:aliases_catalog"),
            {
                "master_dup_scope": "insumos",
                "master_dup_min_count": "2",
                "export": "master_duplicates_csv",
            },
        )
        self.assertEqual(response_csv.status_code, 200)
        self.assertIn("text/csv", response_csv["Content-Type"])
        body = response_csv.content.decode("utf-8")
        self.assertIn("group_type,duplicate_key,count,model,id,nombre,activo,codigo_point", body)
        self.assertIn("mantequilla barra", body)

        response_xlsx = self.client.get(
            reverse("inventario:aliases_catalog"),
            {
                "master_dup_scope": "insumos",
                "master_dup_min_count": "2",
                "export": "master_duplicates_xlsx",
            },
        )
        self.assertEqual(response_xlsx.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response_xlsx["Content-Type"],
        )
        wb = load_workbook(BytesIO(response_xlsx.content), data_only=True)
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, 9)]
        self.assertEqual(
            headers,
            ["group_type", "duplicate_key", "count", "model", "id", "nombre", "activo", "codigo_point"],
        )

    def test_resolve_duplicate_insumo_merges_relations_and_deactivates_source(self):
        unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA)
        proveedor = Proveedor.objects.create(nombre="Proveedor Merge")
        source = Insumo.objects.create(nombre="Mantequilla Barra", unidad_base=unidad, proveedor_principal=proveedor)
        target = Insumo.objects.create(nombre="MANTEQUILLA BARRA", unidad_base=unidad)
        alias = InsumoAlias.objects.create(nombre="Mantequilla barra premium", insumo=source)
        CostoInsumo.objects.create(
            insumo=source,
            proveedor=proveedor,
            costo_unitario=Decimal("120.000000"),
            source_hash="hash-merge-costo-source-001",
        )
        receta = Receta.objects.create(nombre="Receta Merge Insumo", hash_contenido="hash-merge-receta-001")
        linea_with_source = LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=source,
            insumo_texto=source.nombre,
            cantidad=Decimal("2"),
            unidad=unidad,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("120.000000"),
            match_status=LineaReceta.STATUS_AUTO,
        )
        linea_pending = LineaReceta.objects.create(
            receta=receta,
            posicion=2,
            insumo=None,
            insumo_texto=source.nombre,
            cantidad=Decimal("1"),
            unidad=None,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("0"),
            match_status=LineaReceta.STATUS_REJECTED,
        )
        MovimientoInventario.objects.create(
            tipo=MovimientoInventario.TIPO_ENTRADA,
            insumo=source,
            cantidad=Decimal("3"),
            referencia="MV-MERGE-001",
            source_hash="hash-mv-merge-001",
        )
        AjusteInventario.objects.create(
            insumo=source,
            cantidad_sistema=Decimal("10"),
            cantidad_fisica=Decimal("9"),
            motivo="Ajuste merge",
            estatus=AjusteInventario.STATUS_PENDIENTE,
            solicitado_por=self.user,
        )
        SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin_inv",
            insumo=source,
            cantidad=Decimal("5"),
        )
        ExistenciaInsumo.objects.create(
            insumo=source,
            stock_actual=Decimal("10"),
            punto_reorden=Decimal("3"),
            stock_minimo=Decimal("4"),
            stock_maximo=Decimal("12"),
            inventario_promedio=Decimal("8"),
            dias_llegada_pedido=2,
            consumo_diario_promedio=Decimal("1"),
        )
        target_ex = ExistenciaInsumo.objects.create(
            insumo=target,
            stock_actual=Decimal("7"),
            punto_reorden=Decimal("4"),
            stock_minimo=Decimal("5"),
            stock_maximo=Decimal("11"),
            inventario_promedio=Decimal("6"),
            dias_llegada_pedido=1,
            consumo_diario_promedio=Decimal("2"),
        )
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-MANTE-001",
            point_nombre=source.nombre,
            fuzzy_score=99.0,
            fuzzy_sugerencia=target.nombre,
        )

        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {
                "action": "resolve_duplicate_insumo",
                "source_insumo_id": str(source.id),
                "target_insumo_id": str(target.id),
            },
        )
        self.assertEqual(response.status_code, 302)

        source.refresh_from_db()
        target.refresh_from_db()
        alias.refresh_from_db()
        linea_with_source.refresh_from_db()
        linea_pending.refresh_from_db()
        target_ex.refresh_from_db()

        self.assertFalse(source.activo)
        self.assertEqual(alias.insumo_id, target.id)
        self.assertEqual(CostoInsumo.objects.filter(insumo=target).count(), 1)
        self.assertEqual(linea_with_source.insumo_id, target.id)
        self.assertEqual(linea_pending.insumo_id, target.id)
        self.assertEqual(linea_pending.match_status, LineaReceta.STATUS_AUTO)
        self.assertEqual(linea_pending.match_method, "ALIAS")
        self.assertEqual(MovimientoInventario.objects.filter(insumo=target).count(), 1)
        self.assertEqual(AjusteInventario.objects.filter(insumo=target).count(), 1)
        self.assertEqual(SolicitudCompra.objects.filter(insumo=target).count(), 1)
        self.assertEqual(target_ex.stock_actual, Decimal("17"))
        self.assertFalse(ExistenciaInsumo.objects.filter(insumo=source).exists())
        self.assertEqual(target.codigo_point, "PT-MANTE-001")
        self.assertEqual(target.nombre_point, source.nombre)
        self.assertFalse(PointPendingMatch.objects.filter(point_codigo="PT-MANTE-001").exists())
        self.assertTrue(
            AuditLog.objects.filter(action="MASTER_DUPLICATE_RESOLVE_INSUMO", model="maestros.Insumo").exists()
        )

    def test_resolve_duplicate_insumo_group_consolidates_members(self):
        unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA)
        proveedor = Proveedor.objects.create(nombre="Proveedor Group")
        insumo_a = Insumo.objects.create(nombre="Azucar Glass", unidad_base=unidad, proveedor_principal=proveedor)
        insumo_b = Insumo.objects.create(nombre="AZUCAR   GLASS", unidad_base=unidad)
        insumo_c = Insumo.objects.create(nombre="Azúcar Glass", unidad_base=unidad)
        for idx, insumo in enumerate([insumo_a, insumo_b, insumo_c], start=1):
            InsumoAlias.objects.create(nombre=f"Alias {idx} Azucar Glass", insumo=insumo)
            CostoInsumo.objects.create(
                insumo=insumo,
                proveedor=proveedor,
                costo_unitario=Decimal("10.000000") + Decimal(idx),
                source_hash=f"hash-merge-group-{idx}",
            )
            ExistenciaInsumo.objects.create(
                insumo=insumo,
                stock_actual=Decimal(idx),
                punto_reorden=Decimal("1"),
                stock_minimo=Decimal("1"),
                stock_maximo=Decimal("3"),
                inventario_promedio=Decimal("1"),
                dias_llegada_pedido=1,
                consumo_diario_promedio=Decimal("1"),
            )

        response = self.client.post(
            reverse("inventario:aliases_catalog"),
            {
                "action": "resolve_duplicate_insumo_group",
                "duplicate_key": "azucar glass",
                "target_insumo_id": str(insumo_b.id),
            },
        )
        self.assertEqual(response.status_code, 302)

        insumo_a.refresh_from_db()
        insumo_b.refresh_from_db()
        insumo_c.refresh_from_db()
        target_ex = ExistenciaInsumo.objects.get(insumo=insumo_b)

        self.assertFalse(insumo_a.activo)
        self.assertTrue(insumo_b.activo)
        self.assertFalse(insumo_c.activo)
        self.assertEqual(InsumoAlias.objects.filter(insumo=insumo_b).count(), 4)
        self.assertEqual(CostoInsumo.objects.filter(insumo=insumo_b).count(), 3)
        self.assertEqual(target_ex.stock_actual, Decimal("6"))
        self.assertFalse(ExistenciaInsumo.objects.filter(insumo=insumo_a).exists())
        self.assertFalse(ExistenciaInsumo.objects.filter(insumo=insumo_c).exists())

        log = AuditLog.objects.filter(action="MASTER_DUPLICATE_RESOLVE_GROUP", model="maestros.Insumo").first()
        self.assertIsNotNone(log)
        self.assertEqual(log.payload.get("target_id"), insumo_b.id)
        self.assertEqual(int(log.payload.get("sources_resolved", 0)), 2)

    def test_inventory_core_views_render_maturity_and_handoff_sections(self):
        for url in (
            reverse("inventario:existencias"),
            reverse("inventario:movimientos"),
            reverse("inventario:ajustes"),
            reverse("inventario:alertas"),
        ):
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200)
            self.assertIn("maturity_summary", response.context)
            self.assertIn("handoff_map", response.context)
            self.assertIn("erp_governance_rows", response.context)
            self.assertIn("downstream_handoff_rows", response.context)
            self.assertIn("executive_radar_rows", response.context)
            self.assertIn("erp_command_center", response.context)

    def test_importar_archivos_view_renders_governance_table(self):
        response = self.client.get(reverse("inventario:importar_archivos"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Workflow ERP de carga de almacén")
        self.assertContains(response, "Radar ejecutivo ERP")
        self.assertContains(response, "Mesa de gobierno ERP")
        self.assertIn("executive_radar_rows", response.context)
        self.assertIn("erp_governance_rows", response.context)

    def test_existencias_view_shows_release_gate_enterprise_block(self):
        response = self.client.get(reverse("inventario:existencias"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Criterios de cierre")
        self.assertContains(response, "Cierre global")
        self.assertIn("release_gate_rows", response.context)
        self.assertIn("release_gate_completion", response.context)


class InventarioAjustesApprovalTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.admin = user_model.objects.create_user(
            username="admin_aprueba_inv",
            email="admin_aprueba_inv@example.com",
            password="test12345",
        )
        self.almacen = user_model.objects.create_user(
            username="almacen_inv",
            email="almacen_inv@example.com",
            password="test12345",
        )
        admin_group, _ = Group.objects.get_or_create(name="ADMIN")
        almacen_group, _ = Group.objects.get_or_create(name="ALMACEN")
        self.admin.groups.add(admin_group)
        self.almacen.groups.add(almacen_group)

        self.unidad = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA)
        self.insumo = Insumo.objects.create(nombre="Azucar Ajuste", unidad_base=self.unidad, activo=True)
        self.existencia = ExistenciaInsumo.objects.create(insumo=self.insumo, stock_actual=Decimal("10"))

    def test_almacen_registra_ajuste_queda_pendiente(self):
        self.client.force_login(self.almacen)
        response = self.client.post(
            reverse("inventario:ajustes"),
            {
                "action": "create",
                "insumo_id": self.insumo.id,
                "cantidad_sistema": "10",
                "cantidad_fisica": "8",
                "motivo": "Conteo semanal",
                "create_and_apply": "1",
            },
        )
        self.assertEqual(response.status_code, 302)

        ajuste = AjusteInventario.objects.latest("id")
        self.assertEqual(ajuste.estatus, AjusteInventario.STATUS_PENDIENTE)
        self.assertEqual(ajuste.solicitado_por_id, self.almacen.id)
        self.existencia.refresh_from_db()
        self.assertEqual(self.existencia.stock_actual, Decimal("10"))
        self.assertFalse(MovimientoInventario.objects.filter(referencia=ajuste.folio).exists())

    def test_admin_aprueba_y_aplica_ajuste(self):
        ajuste = AjusteInventario.objects.create(
            insumo=self.insumo,
            cantidad_sistema=Decimal("10"),
            cantidad_fisica=Decimal("8"),
            motivo="Conteo mensual",
            estatus=AjusteInventario.STATUS_PENDIENTE,
            solicitado_por=self.almacen,
        )

        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("inventario:ajustes"),
            {
                "action": "approve",
                "ajuste_id": ajuste.id,
                "comentario_revision": "Aprobado por diferencias de merma",
            },
        )
        self.assertEqual(response.status_code, 302)

        ajuste.refresh_from_db()
        self.assertEqual(ajuste.estatus, AjusteInventario.STATUS_APLICADO)
        self.assertEqual(ajuste.aprobado_por_id, self.admin.id)
        self.assertIsNotNone(ajuste.aprobado_en)
        self.assertIsNotNone(ajuste.aplicado_en)

        self.existencia.refresh_from_db()
        self.assertEqual(self.existencia.stock_actual, Decimal("8"))
        movimiento = MovimientoInventario.objects.filter(referencia=ajuste.folio).first()
        self.assertIsNotNone(movimiento)
        self.assertEqual(movimiento.cantidad, Decimal("2"))
        self.assertEqual(movimiento.tipo, MovimientoInventario.TIPO_SALIDA)

    def test_admin_aprueba_ajuste_de_variante_y_aplica_en_canonico(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor Ajuste Canon", activo=True)
        unidad = UnidadMedida.objects.create(
            codigo="kg-aj-can",
            nombre="Kilogramo Ajuste Canon",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        canonical = Insumo.objects.create(
            nombre="Azucar Canonica Ajuste",
            categoria="Masa",
            unidad_base=unidad,
            proveedor_principal=proveedor,
            activo=True,
            codigo_point="AJ-CAN-001",
        )
        variant = Insumo.objects.create(
            nombre="AZUCAR CANONICA AJUSTE",
            categoria="Masa",
            unidad_base=unidad,
            activo=True,
        )
        ExistenciaInsumo.objects.create(insumo=canonical, stock_actual=Decimal("10"))
        ExistenciaInsumo.objects.create(insumo=variant, stock_actual=Decimal("6"))
        ajuste = AjusteInventario.objects.create(
            insumo=variant,
            cantidad_sistema=Decimal("6"),
            cantidad_fisica=Decimal("4"),
            motivo="Conteo canonico",
            estatus=AjusteInventario.STATUS_PENDIENTE,
            solicitado_por=self.almacen,
        )

        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("inventario:ajustes"),
            {
                "action": "approve",
                "ajuste_id": ajuste.id,
                "comentario_revision": "Aprobado usando canonico",
            },
        )
        self.assertEqual(response.status_code, 302)

        canonical_ex = ExistenciaInsumo.objects.get(insumo=canonical)
        variant_ex = ExistenciaInsumo.objects.get(insumo=variant)
        self.assertEqual(canonical_ex.stock_actual, Decimal("8"))
        self.assertEqual(variant_ex.stock_actual, Decimal("6"))
        movimiento = MovimientoInventario.objects.filter(referencia=ajuste.folio).first()
        self.assertIsNotNone(movimiento)
        self.assertEqual(movimiento.insumo_id, canonical.id)

    def test_admin_rechaza_ajuste_sin_afectar_stock(self):
        ajuste = AjusteInventario.objects.create(
            insumo=self.insumo,
            cantidad_sistema=Decimal("10"),
            cantidad_fisica=Decimal("7"),
            motivo="Conteo de cierre",
            estatus=AjusteInventario.STATUS_PENDIENTE,
            solicitado_por=self.almacen,
        )

        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("inventario:ajustes"),
            {
                "action": "reject",
                "ajuste_id": ajuste.id,
                "comentario_revision": "Falta evidencia de conteo físico",
            },
        )
        self.assertEqual(response.status_code, 302)

        ajuste.refresh_from_db()
        self.assertEqual(ajuste.estatus, AjusteInventario.STATUS_RECHAZADO)
        self.assertEqual(ajuste.aprobado_por_id, self.admin.id)
        self.existencia.refresh_from_db()
        self.assertEqual(self.existencia.stock_actual, Decimal("10"))
        self.assertFalse(MovimientoInventario.objects.filter(referencia=ajuste.folio).exists())

    def test_almacen_no_puede_aprobar_ajuste(self):
        ajuste = AjusteInventario.objects.create(
            insumo=self.insumo,
            cantidad_sistema=Decimal("10"),
            cantidad_fisica=Decimal("9"),
            motivo="Conteo rápido",
            estatus=AjusteInventario.STATUS_PENDIENTE,
            solicitado_por=self.almacen,
        )
        self.client.force_login(self.almacen)
        response = self.client.post(
            reverse("inventario:ajustes"),
            {"action": "approve", "ajuste_id": ajuste.id},
        )
        self.assertEqual(response.status_code, 403)
