from io import BytesIO
from decimal import Decimal
from datetime import timedelta

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from core.models import AuditLog, Sucursal
from compras.models import SolicitudCompra
from inventario.models import AjusteInventario, ExistenciaInsumo, MovimientoInventario
from maestros.models import CostoInsumo, Insumo, InsumoAlias, PointPendingMatch, Proveedor, UnidadMedida
from maestros.utils.canonical_catalog import canonical_member_ids, latest_costo_canonico
from recetas.models import LineaReceta, Receta, VentaHistorica


class PointPendingReviewTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_maestros",
            email="admin_maestros@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

        self.unidad = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.proveedor = Proveedor.objects.create(nombre="Proveedor Test", activo=True)
        self.insumo_harina = Insumo.objects.create(
            nombre="Harina Pastelera",
            categoria="Masa",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor,
            activo=True,
        )

    def test_point_pending_export_csv_applies_filters(self):
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-INS-001",
            point_nombre="Harina Point",
            fuzzy_sugerencia="Harina Pastelera",
            fuzzy_score=95.0,
            method="FUZZY",
        )
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-INS-002",
            point_nombre="Azucar Point",
            fuzzy_sugerencia="Azucar",
            fuzzy_score=40.0,
            method="FUZZY",
        )
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_PROVEEDOR,
            point_codigo="PT-PROV-001",
            point_nombre="Proveedor Point",
            fuzzy_sugerencia="Proveedor Test",
            fuzzy_score=99.0,
            method="EXACT",
        )

        response = self.client.get(
            reverse("maestros:point_pending_review"),
            {
                "tipo": "INSUMO",
                "q": "Harina",
                "score_min": "90",
                "export": "csv",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        body = response.content.decode("utf-8")
        self.assertIn("PT-INS-001", body)
        self.assertNotIn("PT-INS-002", body)
        self.assertNotIn("PT-PROV-001", body)
        self.assertIn("INSUMO,Harina,90.0,1", body)

    def test_point_pending_export_xlsx_includes_headers(self):
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-INS-003",
            point_nombre="Mantequilla Point",
            fuzzy_sugerencia="Mantequilla",
            fuzzy_score=92.5,
            method="FUZZY",
        )

        response = self.client.get(
            reverse("maestros:point_pending_review"),
            {"tipo": "INSUMO", "export": "xlsx"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )

        wb = load_workbook(BytesIO(response.content), data_only=True)
        ws = wb.active
        headers = [ws.cell(row=4, column=i).value for i in range(1, 9)]
        self.assertEqual(
            headers,
            [
                "id",
                "tipo",
                "codigo_point",
                "nombre_point",
                "sugerencia",
                "score",
                "metodo",
                "creado_en",
            ],
        )
        self.assertEqual(ws.cell(row=5, column=3).value, "PT-INS-003")

    def test_point_pending_review_hides_historical_records_by_default(self):
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-HIST-001",
            point_nombre="Articulo Historico",
            method="POINT_BRIDGE_MOVEMENTS",
            clasificacion_operativa=PointPendingMatch.CLASIFICACION_OPERATIVA_HISTORICO,
            visible_en_operacion=False,
        )
        visible = PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-ACT-001",
            point_nombre="Articulo Activo",
            method="FUZZY",
        )

        response = self.client.get(reverse("maestros:point_pending_review"), {"tipo": "INSUMO"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context["page"].object_list), [visible])
        self.assertEqual(response.context["counts"][PointPendingMatch.TIPO_INSUMO], 1)

    def test_point_pending_review_search_includes_hidden_historical_records(self):
        hidden = PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-HIST-002",
            point_nombre="Rosca Temporada Especial",
            method="POINT_BRIDGE_MOVEMENTS",
            clasificacion_operativa=PointPendingMatch.CLASIFICACION_OPERATIVA_HISTORICO,
            visible_en_operacion=False,
        )

        response = self.client.get(
            reverse("maestros:point_pending_review"),
            {"tipo": "INSUMO", "q": "Rosca"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context["page"].object_list), [hidden])

    def test_auto_resolve_sugerencias_insumos_updates_mapping_and_alias(self):
        pending = PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-INS-010",
            point_nombre="Harina Point Oficial",
            fuzzy_sugerencia="Harina Pastelera",
            fuzzy_score=97.0,
            method="FUZZY",
        )

        response = self.client.post(
            reverse("maestros:point_pending_review"),
            {
                "tipo": "INSUMO",
                "action": "resolve_sugerencias_insumos",
                "pending_ids": [str(pending.id)],
                "auto_score_min": "90",
                "create_aliases": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(PointPendingMatch.objects.filter(id=pending.id).exists())

        self.insumo_harina.refresh_from_db()
        self.assertEqual(self.insumo_harina.codigo_point, "PT-INS-010")
        self.assertEqual(self.insumo_harina.nombre_point, "Harina Point Oficial")

        alias_norm = "harina point oficial"
        self.assertTrue(
            InsumoAlias.objects.filter(nombre_normalizado=alias_norm, insumo=self.insumo_harina).exists()
        )

    def test_auto_resolve_sugerencias_uses_current_filter_when_no_selection(self):
        pending_ok = PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-INS-020",
            point_nombre="Harina Point Filtrada",
            fuzzy_sugerencia="Harina Pastelera",
            fuzzy_score=95.0,
            method="FUZZY",
        )
        pending_other = PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-INS-021",
            point_nombre="Azucar Point",
            fuzzy_sugerencia="Azucar",
            fuzzy_score=99.0,
            method="FUZZY",
        )

        response = self.client.post(
            reverse("maestros:point_pending_review"),
            {
                "tipo": "INSUMO",
                "action": "resolve_sugerencias_insumos",
                "q": "Harina",
                "score_min": "90",
                "auto_score_min": "90",
                "create_aliases": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(PointPendingMatch.objects.filter(id=pending_ok.id).exists())
        self.assertTrue(PointPendingMatch.objects.filter(id=pending_other.id).exists())

    def test_point_pending_review_insumo_selector_uses_canonicalized_options(self):
        canonical = Insumo.objects.create(
            nombre="Etiqueta Canonica",
            codigo="ETQ-CAN-01",
            categoria="Etiquetas",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        Insumo.objects.create(
            nombre="ETIQUETA CANONICA",
            categoria="Etiquetas",
            unidad_base=self.unidad,
            activo=True,
        )

        response = self.client.get(reverse("maestros:point_pending_review"), {"tipo": "INSUMO"})
        self.assertEqual(response.status_code, 200)

        insumos = list(response.context["insumos"])
        names = [item.nombre for item in insumos]
        self.assertIn("Etiqueta Canonica", names)
        self.assertNotIn("ETIQUETA CANONICA", names)

        canonical_option = next(item for item in insumos if item.nombre == "Etiqueta Canonica")
        self.assertEqual(canonical_option.canonical_variant_count, 2)

    def test_point_pending_review_shows_canonical_target_for_insumo_suggestion(self):
        canonical = Insumo.objects.create(
            nombre="Etiqueta Canonica",
            codigo="ETQ-CAN-02",
            categoria="Etiquetas",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        Insumo.objects.create(
            nombre="ETIQUETA CANONICA",
            categoria="Etiquetas",
            unidad_base=self.unidad,
            activo=True,
        )
        pending = PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-ETQ-022",
            point_nombre="Etiqueta Point",
            fuzzy_sugerencia="ETIQUETA CANONICA",
            fuzzy_score=98.0,
            method="EXACT",
        )

        response = self.client.get(reverse("maestros:point_pending_review"), {"tipo": "INSUMO"})
        self.assertEqual(response.status_code, 200)
        row = next(item for item in response.context["page"].object_list if item.id == pending.id)
        self.assertEqual(row.canonical_target.id, canonical.id)
        self.assertEqual(row.canonical_target_variants, 2)
        self.assertContains(response, "Artículo maestro propuesto:")
        self.assertContains(response, "Etiqueta Canonica")

    def test_point_pending_review_renders_enterprise_workflow(self):
        PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-INS-WF-01",
            point_nombre="Chocolate Comercial",
            fuzzy_sugerencia="Chocolate",
            fuzzy_score=91.0,
            method="FUZZY",
        )

        response = self.client.get(reverse("maestros:point_pending_review"), {"tipo": "INSUMO"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Centro de mando ERP")
        self.assertContains(response, "Workflow ERP de integración comercial")
        self.assertContains(response, "Ruta crítica ERP")
        self.assertContains(response, "Radar ejecutivo ERP")
        self.assertContains(response, "Mesa de gobierno ERP")
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)

    def test_point_pending_resolve_insumos_normaliza_variante_al_canonico(self):
        canonical = Insumo.objects.create(
            nombre="Etiqueta Canonica Resolver",
            codigo="ETQ-CAN-03",
            categoria="Etiquetas",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        variant = Insumo.objects.create(
            nombre="ETIQUETA CANONICA RESOLVER",
            categoria="Etiquetas",
            unidad_base=self.unidad,
            activo=True,
        )
        pending = PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-ETQ-099",
            point_nombre="Etiqueta Point Resolver",
            fuzzy_sugerencia="ETIQUETA CANONICA RESOLVER",
            fuzzy_score=97.0,
            method="FUZZY",
        )

        response = self.client.post(
            reverse("maestros:point_pending_review"),
            {
                "tipo": "INSUMO",
                "pending_ids": [str(pending.id)],
                "action": "resolve_insumos",
                "insumo_id": str(variant.id),
                "create_aliases": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertFalse(PointPendingMatch.objects.filter(id=pending.id).exists())
        canonical.refresh_from_db()
        self.assertEqual(canonical.codigo_point, "PT-ETQ-099")

    def test_point_pending_post_redirect_preserves_filters(self):
        pending = PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-INS-099",
            point_nombre="Harina Point Redirect",
            fuzzy_sugerencia="Harina Pastelera",
            fuzzy_score=95.0,
            method="FUZZY",
        )

        response = self.client.post(
            reverse("maestros:point_pending_review"),
            {
                "tipo": "INSUMO",
                "action": "resolve_sugerencias_insumos",
                "pending_ids": [str(pending.id)],
                "q": "Harina",
                "score_min": "90",
                "page": "3",
                "auto_score_min": "90",
                "create_aliases": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("tipo=INSUMO", response.url)
        self.assertIn("q=Harina", response.url)
        self.assertIn("score_min=90.0", response.url)
        self.assertIn("page=3", response.url)


class ProveedorListEnterpriseTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_proveedores_erp",
            email="admin_proveedores_erp@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)
        Proveedor.objects.create(nombre="Proveedor Activo ERP", lead_time_dias=3, activo=True)
        Proveedor.objects.create(nombre="Proveedor Inactivo ERP", lead_time_dias=7, activo=False)

    def test_proveedor_list_renders_enterprise_cockpit(self):
        response = self.client.get(reverse("maestros:proveedor_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Centro de mando ERP")
        self.assertContains(response, "Workflow ERP del proveedor")
        self.assertContains(response, "Ruta crítica ERP")
        self.assertContains(response, "Radar ejecutivo ERP")
        self.assertContains(response, "Mesa de gobierno ERP")
        self.assertContains(response, "Lead time promedio")
        self.assertContains(response, "Inactivos")
        self.assertIn("erp_command_center", response.context)
        self.assertIn("workflow_rows", response.context)
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertIn("erp_governance_rows", response.context)

    def test_proveedor_create_form_renders_enterprise_cockpit(self):
        response = self.client.get(reverse("maestros:proveedor_create"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Centro de mando ERP")
        self.assertContains(response, "Workflow ERP del proveedor")
        self.assertContains(response, "Ruta crítica ERP")
        self.assertContains(response, "Radar ejecutivo ERP")
        self.assertContains(response, "Checklist de cierre ERP")
        self.assertContains(response, "Documento ERP activo")
        self.assertIn("erp_command_center", response.context)
        self.assertIn("workflow_rows", response.context)
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertIn("provider_checklist", response.context)
        self.assertEqual(response.context["provider_form_mode"], "alta")

    def test_proveedor_update_form_renders_enterprise_cockpit(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor Form ERP", lead_time_dias=5, activo=True)
        response = self.client.get(reverse("maestros:proveedor_update", args=[proveedor.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Centro de mando ERP")
        self.assertContains(response, "Workflow ERP del proveedor")
        self.assertContains(response, "Ruta crítica ERP")
        self.assertContains(response, "Radar ejecutivo ERP")
        self.assertContains(response, "Checklist de cierre ERP")
        self.assertContains(response, "Documento ERP activo")
        self.assertContains(response, "Proveedor Form ERP")
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertEqual(response.context["provider_form_mode"], "edición")

    def test_proveedor_delete_form_renders_enterprise_cockpit(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor Baja ERP", lead_time_dias=4, activo=True)
        response = self.client.get(reverse("maestros:proveedor_delete", args=[proveedor.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Centro de mando ERP")
        self.assertContains(response, "Workflow ERP del retiro")
        self.assertContains(response, "Ruta crítica ERP")
        self.assertContains(response, "Radar ejecutivo ERP")
        self.assertContains(response, "Checklist de baja ERP")
        self.assertContains(response, "Retirar proveedor del maestro")
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertIn("provider_delete_checklist", response.context)


class InsumoTipoItemCatalogTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_insumo_tipo",
            email="admin_insumo_tipo@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)
        self.unidad = UnidadMedida.objects.create(
            codigo="pza",
            nombre="Pieza",
            tipo=UnidadMedida.TIPO_PIEZA,
            factor_to_base=Decimal("1"),
        )
        self.insumo_mp = Insumo.objects.create(
            nombre="Harina Selecta",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            categoria="Masa",
            unidad_base=self.unidad,
            activo=True,
        )
        self.insumo_int = Insumo.objects.create(
            nombre="Batida Vainilla",
            tipo_item=Insumo.TIPO_INTERNO,
            categoria="Batidas",
            unidad_base=self.unidad,
            activo=True,
        )
        self.insumo_emp = Insumo.objects.create(
            nombre="Caja Pastel M",
            tipo_item=Insumo.TIPO_EMPAQUE,
            categoria="Empaque pastel",
            unidad_base=self.unidad,
            activo=True,
        )

    def test_insumo_list_filters_by_tipo_item(self):
        response = self.client.get(reverse("maestros:insumo_list"), {"tipo_item": Insumo.TIPO_INTERNO})
        self.assertEqual(response.status_code, 200)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertIn(self.insumo_int.nombre, names)
        self.assertNotIn(self.insumo_mp.nombre, names)
        self.assertNotIn(self.insumo_emp.nombre, names)

    def test_insumo_list_context_includes_tipo_item_kpis(self):
        response = self.client.get(reverse("maestros:insumo_list"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_materia_prima"], 1)
        self.assertEqual(response.context["total_insumos_internos"], 1)
        self.assertEqual(response.context["total_empaques"], 1)

    def test_insumo_list_shows_enterprise_document_chain(self):
        response = self.client.get(reverse("maestros:insumo_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Centro de mando ERP")
        self.assertContains(response, "Cadena documental ERP")
        self.assertContains(response, "Ruta crítica ERP")
        self.assertContains(response, "Modelo SAP / ERP del artículo")
        self.assertContains(response, "Cierre por etapa documental")
        self.assertContains(response, "Mesa de gobierno ERP")
        self.assertContains(response, "Frente")
        self.assertContains(response, "Cierre global:")
        self.assertContains(response, "Salud operativa ERP")
        self.assertIn("enterprise_chain", response.context)
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("erp_model_rows", response.context)
        self.assertIn("erp_model_completion", response.context)
        self.assertIn("document_stage_rows", response.context)
        self.assertIn("erp_governance_rows", response.context)
        self.assertTrue(response.context["operational_health_cards"])

    def test_insumo_list_context_includes_recipe_hierarchy_counts(self):
        Receta.objects.create(
            nombre="Batida Base QA",
            hash_contenido="hash-maestro-jerarquia-001",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=False,
        )
        Receta.objects.create(
            nombre="Batida Derivada QA",
            hash_contenido="hash-maestro-jerarquia-002",
            tipo=Receta.TIPO_PREPARACION,
            usa_presentaciones=True,
        )
        Receta.objects.create(
            nombre="Pastel QA Final",
            hash_contenido="hash-maestro-jerarquia-003",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        response = self.client.get(reverse("maestros:insumo_list"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_productos_finales"], 1)
        self.assertEqual(response.context["total_batidas_base"], 1)
        self.assertEqual(response.context["total_subinsumos_derivados"], 1)
        self.assertContains(response, "Jerarquía enterprise del artículo")
        self.assertContains(response, "Producto final")
        self.assertContains(response, reverse("maestros:insumo_create") + "?tipo_item=MATERIA_PRIMA")
        self.assertContains(response, reverse("maestros:insumo_create") + "?tipo_item=INSUMO_INTERNO")
        self.assertContains(response, reverse("maestros:insumo_create") + "?tipo_item=EMPAQUE")
        self.assertContains(response, reverse("recetas:receta_create") + "?mode=BASE")
        self.assertContains(response, reverse("recetas:receta_create") + "?mode=BASE_DERIVADOS")

    def test_insumo_list_filters_by_categoria(self):
        response = self.client.get(reverse("maestros:insumo_list"), {"categoria": "Batidas"})
        self.assertEqual(response.status_code, 200)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertEqual(names, [self.insumo_int.nombre])
        self.assertIn("Batidas", response.context["categorias_catalogo"])

    def test_insumo_create_prefills_tipo_and_nombre_from_querystring(self):
        response = self.client.get(
            reverse("maestros:insumo_create"),
            {"tipo_item": Insumo.TIPO_INTERNO, "nombre": "Dream Whip QA"},
        )
        self.assertEqual(response.status_code, 200)
        form = response.context["form"]
        self.assertEqual(form.initial["tipo_item"], Insumo.TIPO_INTERNO)
        self.assertEqual(form.initial["nombre"], "Dream Whip QA")
        self.assertContains(response, "Estado maestro")
        self.assertContains(response, "Batidas")
        self.assertContains(response, "Panes")

    def test_insumo_create_guides_copy_by_selected_class(self):
        response = self.client.get(reverse("maestros:insumo_create"), {"tipo_item": Insumo.TIPO_EMPAQUE})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nueva empaque")
        self.assertContains(response, "Alta guiada:")
        self.assertContains(response, "Material de presentación final")
        self.assertContains(response, "Guardar empaque")
        self.assertContains(response, "Requisitos ERP por clase")
        self.assertContains(response, "Categorías sugeridas")
        self.assertContains(response, "Uso en producto final o presentación")
        self.assertContains(response, "Caja pastel")

    def test_insumo_create_shows_live_erp_readiness_checklist(self):
        response = self.client.get(reverse("maestros:insumo_create"), {"tipo_item": Insumo.TIPO_MATERIA_PRIMA})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Checklist ERP en vivo")
        self.assertContains(response, "Bloqueos actuales")
        self.assertContains(response, "Proveedor principal")
        self.assertContains(response, "Código de venta recomendado")
        self.assertContains(response, "erp-type-requirements-data")
        self.assertContains(response, "erp-type-titles-data")

    def test_insumo_update_shows_live_erp_readiness_checklist(self):
        response = self.client.get(reverse("maestros:insumo_update", args=[self.insumo_mp.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Checklist ERP en vivo")
        self.assertContains(response, "Bloqueos actuales")
        self.assertContains(response, "Falta:")
        self.assertContains(response, "Producto final")
        self.assertContains(response, "Costeo / MRP")
        self.assertContains(response, "Compras")
        self.assertContains(response, "Inventario")

    def test_insumo_delete_form_renders_enterprise_cockpit(self):
        response = self.client.get(reverse("maestros:insumo_delete", args=[self.insumo_int.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Centro de mando ERP")
        self.assertContains(response, "Workflow ERP del retiro")
        self.assertContains(response, "Checklist de baja ERP")
        self.assertContains(response, "Retirar artículo del maestro")
        self.assertIn("delete_checklist", response.context)

    def test_insumo_create_normalizes_codigo_point_on_save(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor Codigo Point QA", activo=True)
        response = self.client.post(
            reverse("maestros:insumo_create"),
            {
                "codigo": " MP-QA-CP-01 ",
                "codigo_point": " pt-demo-001 ",
                "nombre": " Harina Codigo Point QA ",
                "nombre_point": " Point Harina QA ",
                "tipo_item": Insumo.TIPO_MATERIA_PRIMA,
                "categoria": " Harinas ",
                "unidad_base": str(self.unidad.id),
                "proveedor_principal": str(proveedor.id),
                "activo": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        creado = Insumo.objects.get(nombre="Harina Codigo Point QA")
        self.assertEqual(creado.codigo, "MP-QA-CP-01")
        self.assertEqual(creado.codigo_point, "PT-DEMO-001")
        self.assertEqual(creado.nombre_point, "Point Harina QA")
        self.assertEqual(creado.categoria, "Harinas")

    def test_insumo_create_rejects_duplicate_normalized_codigo_point(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor Duplicate Point QA", activo=True)
        Insumo.objects.create(
            nombre="Harina Canonica Point QA",
            codigo_point="PT-DUP-001",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            categoria="Harinas",
            unidad_base=self.unidad,
            proveedor_principal=proveedor,
            activo=True,
        )
        response = self.client.post(
            reverse("maestros:insumo_create"),
            {
                "codigo": "MP-QA-DUP-02",
                "codigo_point": "pt dup 001",
                "nombre": "Harina Duplicada Point QA",
                "nombre_point": "",
                "tipo_item": Insumo.TIPO_MATERIA_PRIMA,
                "categoria": "Harinas",
                "unidad_base": str(self.unidad.id),
                "proveedor_principal": str(proveedor.id),
                "activo": "on",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "ya está ligado a otro artículo activo")
        self.assertFalse(Insumo.objects.filter(nombre="Harina Duplicada Point QA").exists())

    def test_insumo_update_shows_operational_impact_summary(self):
        receta_base = Receta.objects.create(
            nombre="Batida QA Impacto",
            hash_contenido="hash-maestro-impacto-001",
            tipo=Receta.TIPO_PREPARACION,
        )
        receta_final = Receta.objects.create(
            nombre="Pastel QA Impacto",
            hash_contenido="hash-maestro-impacto-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        LineaReceta.objects.create(receta=receta_base, posicion=1, insumo=self.insumo_mp, cantidad=Decimal("1.000000"))
        LineaReceta.objects.create(receta=receta_final, posicion=1, insumo=self.insumo_mp, cantidad=Decimal("2.000000"))
        SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=self.insumo_mp,
            cantidad=Decimal("3.000000"),
        )
        ExistenciaInsumo.objects.create(insumo=self.insumo_mp, stock_actual=Decimal("5.000000"))

        response = self.client.get(reverse("maestros:insumo_update", args=[self.insumo_mp.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Impacto operativo")
        self.assertContains(response, "Navegación rápida")
        self.assertContains(response, "Ver recetas")
        self.assertContains(response, "Ver finales")
        self.assertContains(response, "Ver compras")
        self.assertContains(response, "Ver inventario")
        self.assertContains(response, "Impacto ERP")
        self.assertContains(response, "Pastel QA Impacto")
        self.assertContains(response, "Batida QA Impacto")

    def test_insumo_list_context_includes_enterprise_readiness_kpis(self):
        response = self.client.get(reverse("maestros:insumo_list"))
        self.assertEqual(response.status_code, 200)
        self.assertIn("total_enterprise_ready", response.context)
        self.assertIn("total_enterprise_incomplete", response.context)
        self.assertContains(response, "Listos ERP")
        self.assertContains(response, "Estado maestro")
        self.assertContains(response, "Estado maestro por tipo")
        self.assertContains(response, "Categorías operativas por tipo")
        self.assertContains(response, "Gobierno del maestro")

    def test_insumo_list_shows_quick_create_by_class(self):
        response = self.client.get(reverse("maestros:insumo_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Alta rápida por clase")
        self.assertContains(response, reverse("maestros:insumo_create") + "?tipo_item=MATERIA_PRIMA")
        self.assertContains(response, reverse("maestros:insumo_create") + "?tipo_item=INSUMO_INTERNO")
        self.assertContains(response, reverse("maestros:insumo_create") + "?tipo_item=EMPAQUE")
        self.assertContains(response, "+ Crear materia prima")
        self.assertContains(response, "+ Crear insumo interno")
        self.assertContains(response, "+ Crear empaque")

    def test_insumo_create_rejects_active_mp_without_supplier(self):
        response = self.client.post(
            reverse("maestros:insumo_create"),
            {
                "codigo": "MP-QA-01",
                "codigo_point": "",
                "nombre": "Azucar QA",
                "nombre_point": "",
                "tipo_item": Insumo.TIPO_MATERIA_PRIMA,
                "categoria": "Endulzantes",
                "unidad_base": str(self.unidad.id),
                "activo": "on",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "proveedor principal")
        self.assertFalse(Insumo.objects.filter(nombre="Azucar QA").exists())

    def test_insumo_create_rejects_active_internal_without_category(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor QA Maestro", activo=True)
        response = self.client.post(
            reverse("maestros:insumo_create"),
            {
                "codigo": "INT-QA-01",
                "codigo_point": "",
                "nombre": "Batida QA Sin Categoria",
                "nombre_point": "",
                "tipo_item": Insumo.TIPO_INTERNO,
                "categoria": "",
                "unidad_base": str(self.unidad.id),
                "proveedor_principal": str(proveedor.id),
                "activo": "on",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "debe tener categoría")
        self.assertFalse(Insumo.objects.filter(nombre="Batida QA Sin Categoria").exists())

    def test_insumo_list_filters_by_canonical_status(self):
        canonical = Insumo.objects.create(
            nombre="Etiqueta QA",
            codigo_point="PT-ETQ-001",
            tipo_item=Insumo.TIPO_EMPAQUE,
            categoria="Empaque pastel",
            unidad_base=self.unidad,
            activo=True,
        )
        variant = Insumo.objects.create(
            nombre="ETIQUETA QA",
            tipo_item=Insumo.TIPO_EMPAQUE,
            categoria="Empaque pastel",
            unidad_base=self.unidad,
            activo=True,
        )

        response = self.client.get(reverse("maestros:insumo_list"), {"canonical_status": "canonicos"})
        self.assertEqual(response.status_code, 200)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertIn(canonical.nombre, names)
        self.assertNotIn(variant.nombre, names)

        response = self.client.get(reverse("maestros:insumo_list"), {"canonical_status": "variantes"})
        self.assertEqual(response.status_code, 200)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertIn(variant.nombre, names)
        self.assertNotIn(canonical.nombre, names)

    def test_insumo_list_context_marks_canonical_and_variant_rows(self):
        canonical = Insumo.objects.create(
            nombre="Domo QA",
            codigo_point="PT-DOMO-001",
            tipo_item=Insumo.TIPO_EMPAQUE,
            categoria="Empaque pastel",
            unidad_base=self.unidad,
            activo=True,
        )
        variant = Insumo.objects.create(
            nombre="DOMO QA",
            tipo_item=Insumo.TIPO_EMPAQUE,
            categoria="Empaque pastel",
            unidad_base=self.unidad,
            activo=True,
        )

        response = self.client.get(reverse("maestros:insumo_list"), {"q": "Domo QA"})
        self.assertEqual(response.status_code, 200)
        page_map = {x.id: x for x in response.context["insumos"]}
        self.assertTrue(page_map[canonical.id].is_canonical_record)
        self.assertFalse(page_map[canonical.id].is_duplicate_variant)
        self.assertFalse(page_map[variant.id].is_canonical_record)
        self.assertTrue(page_map[variant.id].is_duplicate_variant)
        self.assertEqual(page_map[variant.id].canonical_target.id, canonical.id)
        self.assertIsNone(page_map[canonical.id].canonical_target)

    def test_insumo_list_filters_by_costo_status_using_canonical_group_cost(self):
        canonical = Insumo.objects.create(
            nombre="Etiqueta Costo QA",
            codigo_point="PT-ETQ-COST-001",
            tipo_item=Insumo.TIPO_EMPAQUE,
            categoria="Empaque pastel",
            unidad_base=self.unidad,
            activo=True,
        )
        variant = Insumo.objects.create(
            nombre="ETIQUETA COSTO QA",
            tipo_item=Insumo.TIPO_EMPAQUE,
            categoria="Empaque pastel",
            unidad_base=self.unidad,
            activo=True,
        )
        CostoInsumo.objects.create(insumo=variant, costo_unitario=Decimal("7.500000"))

        response = self.client.get(reverse("maestros:insumo_list"), {"costo_status": "con_costo", "q": "Etiqueta Costo QA"})
        self.assertEqual(response.status_code, 200)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertIn(canonical.nombre, names)
        self.assertIn(variant.nombre, names)

        response = self.client.get(reverse("maestros:insumo_list"), {"costo_status": "sin_costo", "q": "Etiqueta Costo QA"})
        self.assertEqual(response.status_code, 200)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertNotIn(canonical.nombre, names)
        self.assertNotIn(variant.nombre, names)

    def test_insumo_list_filters_by_enterprise_status(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor ERP QA", activo=True)
        listo = Insumo.objects.create(
            nombre="Harina Lista QA",
            codigo_point="PT-HAR-001",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            categoria="Harinas",
            unidad_base=self.unidad,
            proveedor_principal=proveedor,
            activo=True,
        )
        incompleto = Insumo.objects.create(
            nombre="Batida Incompleta QA",
            tipo_item=Insumo.TIPO_INTERNO,
            categoria="",
            unidad_base=self.unidad,
            activo=True,
        )
        inactivo = Insumo.objects.create(
            nombre="Caja Inactiva QA",
            tipo_item=Insumo.TIPO_EMPAQUE,
            categoria="Caja pastel",
            unidad_base=self.unidad,
            activo=False,
        )

        response = self.client.get(reverse("maestros:insumo_list"), {"enterprise_status": "listos", "q": "QA"})
        self.assertEqual(response.status_code, 200)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertIn(listo.nombre, names)
        self.assertNotIn(incompleto.nombre, names)
        self.assertNotIn(inactivo.nombre, names)

        response = self.client.get(reverse("maestros:insumo_list"), {"enterprise_status": "incompletos", "q": "QA"})
        self.assertEqual(response.status_code, 200)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertIn(incompleto.nombre, names)
        self.assertNotIn(listo.nombre, names)
        self.assertNotIn(inactivo.nombre, names)

        response = self.client.get(reverse("maestros:insumo_list"), {"enterprise_status": "inactivos", "q": "QA"})
        self.assertEqual(response.status_code, 200)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertIn(inactivo.nombre, names)
        self.assertNotIn(listo.nombre, names)
        self.assertNotIn(incompleto.nombre, names)

    def test_insumo_list_context_includes_category_navigation(self):
        Insumo.objects.create(
            nombre="Caja Pastel QA Categoria",
            tipo_item=Insumo.TIPO_EMPAQUE,
            categoria="Caja pastel",
            unidad_base=self.unidad,
            activo=True,
        )
        response = self.client.get(reverse("maestros:insumo_list"))
        self.assertEqual(response.status_code, 200)
        category_rows = response.context["category_navigation"]
        self.assertTrue(any(row["categoria"] == "Caja pastel" and row["tipo_item"] == Insumo.TIPO_EMPAQUE for row in category_rows))
        self.assertContains(response, "Abrir categoría")
        self.assertContains(response, "Ver faltantes")

    def test_insumo_list_filters_by_missing_field(self):
        proveedor = Proveedor.objects.create(nombre="Proveedor Faltantes QA", activo=True)
        sin_unidad = Insumo.objects.create(
            nombre="Articulo Sin Unidad QA",
            tipo_item=Insumo.TIPO_EMPAQUE,
            categoria="Caja pastel",
            activo=True,
        )
        sin_proveedor = Insumo.objects.create(
            nombre="MP Sin Proveedor QA",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            categoria="Harinas",
            unidad_base=self.unidad,
            activo=True,
        )
        listo = Insumo.objects.create(
            nombre="Articulo Listo QA",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            categoria="Harinas",
            unidad_base=self.unidad,
            proveedor_principal=proveedor,
            codigo_point="PT-LISTO-QA",
            activo=True,
        )

        response = self.client.get(reverse("maestros:insumo_list"), {"missing_field": "unidad", "q": "QA"})
        self.assertEqual(response.status_code, 200)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertIn(sin_unidad.nombre, names)
        self.assertNotIn(sin_proveedor.nombre, names)
        self.assertNotIn(listo.nombre, names)

        response = self.client.get(reverse("maestros:insumo_list"), {"missing_field": "proveedor", "q": "QA"})
        self.assertEqual(response.status_code, 200)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertIn(sin_proveedor.nombre, names)
        self.assertNotIn(sin_unidad.nombre, names)
        self.assertNotIn(listo.nombre, names)

    def test_insumo_list_context_includes_usage_navigation_and_usage_profile(self):
        receta = Receta.objects.create(
            nombre="Pastel Uso QA",
            hash_contenido="hash-uso-maestro-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=self.insumo_int,
            insumo_texto=self.insumo_int.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="pza",
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100.0,
        )
        SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=self.insumo_mp,
            cantidad=Decimal("2.000"),
        )
        ExistenciaInsumo.objects.create(
            insumo=self.insumo_emp,
            stock_actual=Decimal("5.000"),
            punto_reorden=Decimal("1.000"),
            stock_minimo=Decimal("1.000"),
            stock_maximo=Decimal("10.000"),
            inventario_promedio=Decimal("3.000"),
            dias_llegada_pedido=2,
            consumo_diario_promedio=Decimal("1.000"),
        )

        response = self.client.get(reverse("maestros:insumo_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Uso operativo del catálogo")
        self.assertContains(response, "En recetas")
        self.assertContains(response, "En compras")
        self.assertContains(response, "En inventario")

        page_map = {x.id: x for x in response.context["insumos"]}
        self.assertTrue(page_map[self.insumo_int.id].usage_profile["used_in_recipes"])
        self.assertEqual(page_map[self.insumo_int.id].usage_profile["final_recipe_count"], 1)
        self.assertTrue(page_map[self.insumo_mp.id].usage_profile["used_in_purchases"])
        self.assertTrue(page_map[self.insumo_emp.id].usage_profile["used_in_inventory"])
        self.assertContains(response, "En producto final")
        self.assertContains(response, "Ver finales")
        self.assertIn("impact_navigation", response.context)
        self.assertEqual(page_map[self.insumo_int.id].impact_profile["level"], "Crítico")
        self.assertTrue(page_map[self.insumo_mp.id].impact_profile["used_in_purchases"])
        self.assertTrue(page_map[self.insumo_emp.id].impact_profile["used_in_inventory"])

    def test_insumo_list_context_includes_impact_navigation_and_render(self):
        receta = Receta.objects.create(
            nombre="Pastel Impacto QA",
            hash_contenido="hash-impacto-maestro-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=self.insumo_int,
            insumo_texto=self.insumo_int.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="pza",
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100.0,
        )
        SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=self.insumo_int,
            cantidad=Decimal("2.000"),
        )
        response = self.client.get(reverse("maestros:insumo_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Impacto operativo enterprise")
        self.assertContains(response, "Multimódulo")
        impact_map = {item["title"]: item["count"] for item in response.context["impact_navigation"]}
        self.assertGreaterEqual(impact_map["En producto final"], 1)
        self.assertGreaterEqual(impact_map["Con compras activas"], 1)
        page_map = {x.id: x for x in response.context["insumos"]}
        self.assertTrue(page_map[self.insumo_int.id].impact_profile["is_multimodule"])
        self.assertContains(response, "Alcance ERP")

    def test_insumo_list_filters_by_impact_scope(self):
        receta = Receta.objects.create(
            nombre="Pastel Impacto Filtro QA",
            hash_contenido="hash-impacto-maestro-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=self.insumo_int,
            insumo_texto=self.insumo_int.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="pza",
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100.0,
        )
        SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=self.insumo_mp,
            cantidad=Decimal("2.000"),
        )
        response = self.client.get(reverse("maestros:insumo_list"), {"impact_scope": "critical"})
        self.assertEqual(response.status_code, 200)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertIn(self.insumo_int.nombre, names)
        self.assertNotIn(self.insumo_mp.nombre, names)

    def test_insumo_list_filters_explicit_enterprise_blockers_by_module(self):
        receta = Receta.objects.create(
            nombre="Pastel Bloqueo MRP QA",
            hash_contenido="hash-impacto-maestro-002b",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=self.insumo_int,
            insumo_texto=self.insumo_int.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="pza",
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100.0,
        )
        SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=self.insumo_mp,
            cantidad=Decimal("2.000"),
        )
        ExistenciaInsumo.objects.create(
            insumo=self.insumo_emp,
            stock_actual=Decimal("5.000"),
            punto_reorden=Decimal("1.000"),
            stock_minimo=Decimal("1.000"),
            stock_maximo=Decimal("10.000"),
            inventario_promedio=Decimal("3.000"),
            dias_llegada_pedido=2,
            consumo_diario_promedio=Decimal("1.000"),
        )

        response_costeo = self.client.get(reverse("maestros:insumo_list"), {"impact_scope": "bloquea_costeo"})
        self.assertEqual(response_costeo.status_code, 200)
        costeo_names = [x.nombre for x in response_costeo.context["insumos"]]
        self.assertIn(self.insumo_int.nombre, costeo_names)
        self.assertNotIn(self.insumo_mp.nombre, costeo_names)

        response_compras = self.client.get(reverse("maestros:insumo_list"), {"impact_scope": "bloquea_compras"})
        self.assertEqual(response_compras.status_code, 200)
        compras_names = [x.nombre for x in response_compras.context["insumos"]]
        self.assertIn(self.insumo_mp.nombre, compras_names)
        self.assertNotIn(self.insumo_int.nombre, compras_names)

        response_inventario = self.client.get(reverse("maestros:insumo_list"), {"impact_scope": "bloquea_inventario"})
        self.assertEqual(response_inventario.status_code, 200)
        inventario_names = [x.nombre for x in response_inventario.context["insumos"]]
        self.assertIn(self.insumo_emp.nombre, inventario_names)
        self.assertNotIn(self.insumo_mp.nombre, inventario_names)

        page_map = {x.id: x for x in response_costeo.context["insumos"]}
        self.assertTrue(page_map[self.insumo_int.id].impact_profile["blocks_costing"])

    def test_insumo_list_can_filter_recipe_usage_scope_finales(self):
        receta_final = Receta.objects.create(
            nombre="Pastel Scope Final QA",
            hash_contenido="hash-uso-maestro-final-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        receta_base = Receta.objects.create(
            nombre="Base Scope QA",
            hash_contenido="hash-uso-maestro-final-002",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("2.0"),
            rendimiento_unidad=self.unidad,
        )
        LineaReceta.objects.create(
            receta=receta_final,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=self.insumo_int,
            insumo_texto=self.insumo_int.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="pza",
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100.0,
        )
        LineaReceta.objects.create(
            receta=receta_base,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=self.insumo_mp,
            insumo_texto=self.insumo_mp.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="pza",
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100.0,
        )

        response = self.client.get(
            reverse("maestros:insumo_list"),
            {"usage_scope": "recipes", "recipe_scope": "finales"},
        )
        self.assertEqual(response.status_code, 200)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertIn(self.insumo_int.nombre, names)
        self.assertNotIn(self.insumo_mp.nombre, names)

    def test_insumo_list_context_includes_operational_blockers_navigation(self):
        receta = Receta.objects.create(
            nombre="Pastel Bloqueo QA",
            hash_contenido="hash-uso-maestro-003",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=self.insumo_int,
            insumo_texto=self.insumo_int.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="pza",
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100.0,
        )
        SolicitudCompra.objects.create(
            area="Compras",
            solicitante="admin",
            insumo=self.insumo_mp,
            cantidad=Decimal("2.000"),
        )
        ExistenciaInsumo.objects.create(
            insumo=self.insumo_emp,
            stock_actual=Decimal("5.000"),
            punto_reorden=Decimal("1.000"),
            stock_minimo=Decimal("1.000"),
            stock_maximo=Decimal("10.000"),
            inventario_promedio=Decimal("3.000"),
            dias_llegada_pedido=2,
            consumo_diario_promedio=Decimal("1.000"),
        )

        response = self.client.get(reverse("maestros:insumo_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Bloqueos operativos")
        self.assertContains(response, "Bloqueos enterprise por módulo")
        self.assertContains(response, "Bloquea recetas")
        self.assertContains(response, "Bloquea producto final")
        self.assertContains(response, "Bloquea compras")
        self.assertContains(response, "Bloquea inventario")
        self.assertContains(response, "Bloquea costeo/MRP")
        self.assertContains(response, "Bloquea operación")
        self.assertContains(response, "Ver bloqueo final")
        self.assertContains(response, "Ver costeo/MRP")
        self.assertContains(response, "Ver bloqueo compras")
        self.assertContains(response, "Ver bloqueo inventario")
        self.assertContains(response, "Bloqueos sobre producto final")
        self.assertContains(response, "Bloqueo final por dato faltante")
        self.assertContains(response, "Sin código comercial")
        self.assertContains(response, "Pastel Bloqueo QA")

        blockers = {item["title"]: item["count"] for item in response.context["operational_blockers_navigation"]}
        self.assertEqual(blockers["Bloquea recetas"], 1)
        self.assertEqual(blockers["Bloquea producto final"], 1)
        self.assertEqual(blockers["Bloquea compras"], 1)
        self.assertEqual(blockers["Bloquea inventario"], 1)

        page_map = {x.id: x for x in response.context["insumos"]}
        self.assertTrue(page_map[self.insumo_int.id].usage_profile["is_operational_blocker"])
        self.assertTrue(page_map[self.insumo_int.id].usage_profile["blocks_costing"])
        self.assertTrue(page_map[self.insumo_int.id].usage_profile["blocks_final_products"])
        self.assertEqual(page_map[self.insumo_int.id].usage_profile["blocking_final_products_count"], 1)
        self.assertIn("Pastel Bloqueo QA", page_map[self.insumo_int.id].usage_profile["final_recipe_examples"])
        self.assertTrue(page_map[self.insumo_mp.id].usage_profile["is_operational_blocker"])
        self.assertTrue(page_map[self.insumo_mp.id].usage_profile["blocks_purchases"])
        self.assertTrue(page_map[self.insumo_emp.id].usage_profile["is_operational_blocker"])
        self.assertTrue(page_map[self.insumo_emp.id].usage_profile["blocks_inventory"])
        self.assertEqual(response.context["final_product_blockers_preview"][0]["nombre"], self.insumo_int.nombre)
        blockers_by_missing = {item["key"]: item["count"] for item in response.context["final_product_blockers_by_missing"]}
        self.assertEqual(blockers_by_missing["codigo_point"], 1)
        self.assertIn("missing_impact_navigation", response.context)
        enterprise_blockers = {item["title"]: item["count"] for item in response.context["enterprise_blocker_navigation"]}
        self.assertEqual(enterprise_blockers["Bloquea producto final"], 1)
        self.assertEqual(enterprise_blockers["Bloquea costeo/MRP"], 1)
        self.assertEqual(enterprise_blockers["Bloquea compras"], 1)
        self.assertEqual(enterprise_blockers["Bloquea inventario"], 1)
        missing_impact_map = {
            (item["missing_key"], item["impact_key"]): item["count"]
            for item in response.context["missing_impact_navigation"]
        }
        self.assertEqual(missing_impact_map[("codigo_point", "critical")], 1)
        self.assertEqual(missing_impact_map[("codigo_point", "compras")], 1)
        self.assertEqual(missing_impact_map[("codigo_point", "inventario")], 1)
        self.assertContains(response, "Faltante + impacto operativo")
        self.assertContains(response, "Código de venta")
        self.assertContains(response, "Bloquea producto final")

    def test_insumo_list_can_filter_by_linked_recipe_id(self):
        receta_objetivo = Receta.objects.create(
            nombre="Pastel Ligado Maestro QA",
            hash_contenido="hash-uso-maestro-linked-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        receta_ajena = Receta.objects.create(
            nombre="Pastel Ligado Maestro AJENO",
            hash_contenido="hash-uso-maestro-linked-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        LineaReceta.objects.create(
            receta=receta_objetivo,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=self.insumo_int,
            insumo_texto=self.insumo_int.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="pza",
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100.0,
        )
        LineaReceta.objects.create(
            receta=receta_ajena,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=self.insumo_mp,
            insumo_texto=self.insumo_mp.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="pza",
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100.0,
        )

        response = self.client.get(
            reverse("maestros:insumo_list"),
            {"usage_scope": "recipes", "linked_recipe_id": receta_objetivo.id},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["linked_recipe"].id, receta_objetivo.id)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertIn(self.insumo_int.nombre, names)
        self.assertNotIn(self.insumo_mp.nombre, names)
        self.assertContains(response, "Vista filtrada a los artículos ligados a la receta")
        self.assertContains(response, receta_objetivo.nombre)
        self.assertContains(response, reverse("recetas:receta_detail", args=[receta_objetivo.id]))

    def test_insumo_list_can_filter_by_exact_insumo_id(self):
        response = self.client.get(
            reverse("maestros:insumo_list"),
            {"insumo_id": self.insumo_int.id},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_insumo"].id, self.insumo_int.id)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertEqual(names, [self.insumo_int.nombre])
        self.assertContains(response, "Vista filtrada al artículo exacto")
        self.assertContains(response, reverse("maestros:insumo_update", args=[self.insumo_int.id]))

    def test_insumo_list_filters_by_usage_scope(self):
        receta = Receta.objects.create(
            nombre="Pastel Filtro Uso QA",
            hash_contenido="hash-uso-maestro-002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=self.insumo_int,
            insumo_texto=self.insumo_int.nombre,
            cantidad=Decimal("1.000000"),
            unidad=self.unidad,
            unidad_texto="pza",
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
            match_score=100.0,
        )

        response = self.client.get(reverse("maestros:insumo_list"), {"usage_scope": "recipes"})
        self.assertEqual(response.status_code, 200)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertIn(self.insumo_int.nombre, names)
        self.assertNotIn(self.insumo_mp.nombre, names)
        self.assertNotIn(self.insumo_emp.nombre, names)

        response = self.client.get(reverse("maestros:insumo_list"), {"usage_scope": "unused"})
        self.assertEqual(response.status_code, 200)
        names = [x.nombre for x in response.context["insumos"]]
        self.assertNotIn(self.insumo_int.nombre, names)
        self.assertIn(self.insumo_mp.nombre, names)
        self.assertIn(self.insumo_emp.nombre, names)


class CanonicalCatalogHelperTests(TestCase):
    def setUp(self):
        self.unidad = UnidadMedida.objects.create(
            codigo="kg-helper",
            nombre="Kilogramo Helper",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.proveedor = Proveedor.objects.create(nombre="Proveedor Helper", activo=True)

    def test_canonical_member_ids_returns_all_active_variants(self):
        canonical = Insumo.objects.create(
            nombre="Ganache Helper",
            codigo_point="PT-HELP-001",
            unidad_base=self.unidad,
            activo=True,
        )
        variant = Insumo.objects.create(
            nombre="GANACHE HELPER",
            unidad_base=self.unidad,
            activo=True,
        )
        self.assertEqual(set(canonical_member_ids(canonical)), {canonical.id, variant.id})
        self.assertEqual(set(canonical_member_ids(insumo_id=variant.id)), {canonical.id, variant.id})

    def test_latest_costo_canonico_uses_latest_cost_from_variant_group(self):
        canonical = Insumo.objects.create(
            nombre="Batida Helper",
            codigo_point="PT-HELP-002",
            unidad_base=self.unidad,
            activo=True,
        )
        variant = Insumo.objects.create(
            nombre="BATIDA HELPER",
            unidad_base=self.unidad,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=variant,
            proveedor=self.proveedor,
            costo_unitario=Decimal("19.250000"),
            source_hash="helper-cost-001",
        )
        self.assertEqual(latest_costo_canonico(canonical), Decimal("19.250000"))
        self.assertEqual(latest_costo_canonico(insumo_id=variant.id), Decimal("19.250000"))


class InsumoCanonicalResolveTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_insumo_canon",
            email="admin_insumo_canon@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)
        self.unidad = UnidadMedida.objects.create(
            codigo="kg",
            nombre="Kilogramo",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.proveedor = Proveedor.objects.create(nombre="Proveedor Canon", activo=True)

    def test_resolve_duplicate_from_maestros_moves_references_to_canonical(self):
        source = Insumo.objects.create(
            nombre="Pan Chocolate",
            codigo="SRC-PAN",
            codigo_point="PT-PAN-001",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        target = Insumo.objects.create(
            nombre="Pan chocolate",
            codigo="TGT-PAN",
            unidad_base=self.unidad,
            activo=True,
        )
        alias = InsumoAlias.objects.create(nombre="Pan choc", insumo=source)
        CostoInsumo.objects.create(
            insumo=source,
            proveedor=self.proveedor,
            costo_unitario=Decimal("12.50"),
            source_hash="canon-source-cost",
        )
        receta = Receta.objects.create(
            nombre="Pastel Canon Test",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="canon-test-hash",
        )
        linea = LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=source,
            insumo_texto=source.nombre,
            cantidad=Decimal("2"),
            unidad=self.unidad,
            unidad_texto="kg",
            match_status=LineaReceta.STATUS_AUTO,
            match_method="MANUAL",
            match_score=100.0,
        )

        response = self.client.post(
            reverse("maestros:insumo_resolve_duplicate"),
            {
                "source_insumo_id": str(source.id),
                "target_insumo_id": str(target.id),
                "next": reverse("maestros:insumo_list"),
            },
        )
        self.assertEqual(response.status_code, 302)

        source.refresh_from_db()
        target.refresh_from_db()
        alias.refresh_from_db()
        linea.refresh_from_db()

        self.assertFalse(source.activo)
        self.assertEqual(alias.insumo_id, target.id)
        self.assertEqual(linea.insumo_id, target.id)
        self.assertEqual(CostoInsumo.objects.filter(insumo=target).count(), 1)
        self.assertEqual(target.codigo_point, "PT-PAN-001")
        self.assertTrue(
            AuditLog.objects.filter(action="MASTER_DUPLICATE_RESOLVE_INSUMO", model="maestros.Insumo").exists()
        )

    def test_resolve_duplicate_group_from_maestros_consolidates_members(self):
        insumo_a = Insumo.objects.create(
            nombre="Azucar Glass",
            codigo="AZ-1",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        insumo_b = Insumo.objects.create(
            nombre="AZUCAR   GLASS",
            codigo="AZ-2",
            codigo_point="PT-AZU-001",
            unidad_base=self.unidad,
            activo=True,
        )
        insumo_c = Insumo.objects.create(
            nombre="Azúcar Glass",
            codigo="AZ-3",
            unidad_base=self.unidad,
            activo=True,
        )

        receta = Receta.objects.create(
            nombre="Receta prueba grupo canon",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="canon-group-test-hash",
        )
        for idx, insumo in enumerate([insumo_a, insumo_b, insumo_c], start=1):
            InsumoAlias.objects.create(nombre=f"Alias grupo {idx}", insumo=insumo)
            CostoInsumo.objects.create(
                insumo=insumo,
                proveedor=self.proveedor,
                costo_unitario=Decimal("10.00") + Decimal(idx),
                source_hash=f"group-source-cost-{idx}",
            )
            ExistenciaInsumo.objects.create(
                insumo=insumo,
                stock_actual=Decimal(idx),
                punto_reorden=Decimal("1"),
                stock_minimo=Decimal("1"),
                stock_maximo=Decimal("3"),
                inventario_promedio=Decimal("1"),
                dias_llegada_pedido=1,
                consumo_diario_promedio=Decimal("1"),
            )
            MovimientoInventario.objects.create(
                insumo=insumo,
                tipo=MovimientoInventario.TIPO_ENTRADA,
                cantidad=Decimal("1"),
                referencia=f"MOV-{idx}",
            )
            AjusteInventario.objects.create(
                insumo=insumo,
                cantidad_sistema=Decimal("5"),
                cantidad_fisica=Decimal("6"),
                motivo=f"Ajuste {idx}",
                solicitado_por=self.user,
            )
            SolicitudCompra.objects.create(
                area="Compras",
                solicitante="admin",
                insumo=insumo,
                cantidad=Decimal("1"),
            )
            LineaReceta.objects.create(
                receta=receta,
                posicion=idx,
                insumo=insumo,
                insumo_texto=insumo.nombre,
                cantidad=Decimal("1"),
                unidad=self.unidad,
                unidad_texto="kg",
                match_status=LineaReceta.STATUS_AUTO,
                match_method="MANUAL",
                match_score=100.0,
            )

        response = self.client.post(
            reverse("maestros:insumo_resolve_duplicate_group"),
            {
                "duplicate_key": "azucar glass",
                "target_insumo_id": str(insumo_b.id),
                "next": reverse("maestros:insumo_list"),
            },
        )
        self.assertEqual(response.status_code, 302)

        insumo_a.refresh_from_db()
        insumo_b.refresh_from_db()
        insumo_c.refresh_from_db()
        target_ex = ExistenciaInsumo.objects.get(insumo=insumo_b)

        self.assertFalse(insumo_a.activo)
        self.assertTrue(insumo_b.activo)
        self.assertFalse(insumo_c.activo)
        self.assertEqual(InsumoAlias.objects.filter(insumo=insumo_b).count(), 4)
        self.assertEqual(CostoInsumo.objects.filter(insumo=insumo_b).count(), 3)
        self.assertEqual(LineaReceta.objects.filter(insumo=insumo_b).count(), 3)
        self.assertEqual(MovimientoInventario.objects.filter(insumo=insumo_b).count(), 3)
        self.assertEqual(AjusteInventario.objects.filter(insumo=insumo_b).count(), 3)
        self.assertEqual(SolicitudCompra.objects.filter(insumo=insumo_b).count(), 3)
        self.assertEqual(target_ex.stock_actual, Decimal("6"))
        self.assertFalse(ExistenciaInsumo.objects.filter(insumo=insumo_a).exists())
        self.assertFalse(ExistenciaInsumo.objects.filter(insumo=insumo_c).exists())
        self.assertTrue(
            AuditLog.objects.filter(action="MASTER_DUPLICATE_RESOLVE_GROUP", model="maestros.Insumo").exists()
        )

    def test_point_pending_manual_resolve_redirects_to_canonical_target(self):
        canonical = Insumo.objects.create(
            nombre="Etiqueta Canonica",
            codigo="ETQ-CANON-01",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor,
            activo=True,
        )
        variant = Insumo.objects.create(
            nombre="ETIQUETA CANONICA",
            unidad_base=self.unidad,
            activo=True,
        )
        pending = PointPendingMatch.objects.create(
            tipo=PointPendingMatch.TIPO_INSUMO,
            point_codigo="PT-ETQ-011",
            point_nombre="Etiqueta Canonica Point",
            fuzzy_sugerencia="Etiqueta Canonica",
            fuzzy_score=99.0,
            method="EXACT",
        )

        response = self.client.post(
            reverse("maestros:point_pending_review"),
            {
                "tipo": "INSUMO",
                "action": "resolve_insumos",
                "pending_ids": [str(pending.id)],
                "insumo_id": str(variant.id),
                "create_aliases": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        canonical.refresh_from_db()
        variant.refresh_from_db()

        self.assertEqual(canonical.nombre_point, "Etiqueta Canonica Point")
        self.assertEqual(canonical.codigo_point, "PT-ETQ-011")
        self.assertTrue(variant.activo)


class MaestroEnterpriseCockpitTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            username="admin_maestro_cockpit",
            email="admin_maestro_cockpit@example.com",
            password="test12345",
        )
        self.client.force_login(self.user)

        self.unidad = UnidadMedida.objects.create(
            codigo="kgc",
            nombre="Kilogramo Cockpit",
            tipo=UnidadMedida.TIPO_MASA,
            factor_to_base=Decimal("1000"),
        )
        self.proveedor = Proveedor.objects.create(nombre="Proveedor Cockpit", activo=True)
        self.insumo = Insumo.objects.create(
            nombre="Chocolate Cockpit",
            categoria="Coberturas",
            unidad_base=self.unidad,
            proveedor_principal=self.proveedor,
            activo=True,
        )

    def test_insumo_list_shows_module_responsables(self):
        response = self.client.get(reverse("maestros:insumo_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Radar ejecutivo ERP")
        self.assertContains(response, "Entrega del maestro a downstream")
        self.assertContains(response, "Ruta troncal ERP")
        self.assertContains(response, "Recetas / BOM")
        self.assertContains(response, "Compras documentales")
        self.assertContains(response, "Inventario / Reabasto")
        self.assertContains(response, "Responsable Producción / Costeo")
        self.assertContains(response, "Responsable Compras")
        self.assertContains(response, "Responsable Inventario / Almacén")
        self.assertIn("executive_radar_rows", response.context)
        self.assertIn("downstream_handoff_rows", response.context)
        self.assertIn("trunk_handoff_rows", response.context)
        self.assertEqual(len(response.context["executive_radar_rows"]), 4)
        self.assertEqual(len(response.context["downstream_handoff_rows"]), 4)
        self.assertEqual(len(response.context["trunk_handoff_rows"]), 3)

    def test_insumo_update_shows_module_responsables_table(self):
        response = self.client.get(reverse("maestros:insumo_update", args=[self.insumo.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Centro de mando ERP")
        self.assertContains(response, "Puerta ERP por módulo")
        self.assertContains(response, "Entrega del maestro a downstream")
        self.assertContains(response, "Ruta troncal ERP")
        self.assertContains(response, "Radar ejecutivo ERP")
        self.assertContains(response, "Ruta crítica ERP del artículo")
        self.assertContains(response, "Cierre multi-módulo del artículo")
        self.assertContains(response, "Mesa de gobierno ERP")
        self.assertContains(response, "Cadena troncal del artículo")
        self.assertContains(response, "Dependencia")
        self.assertContains(response, "<th>Responsable</th>", html=False)
        self.assertContains(response, "Producción / Costeo")
        self.assertContains(response, "Compras")
        self.assertContains(response, "Inventario / Almacén")
        self.assertIn("erp_governance_rows", response.context)
        self.assertIn("downstream_handoff_rows", response.context)
        self.assertIn("trunk_handoff_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertIn("erp_critical_path_rows", response.context)
        self.assertIn("erp_release_rows", response.context)
        self.assertIn("erp_article_chain_rows", response.context)
        self.assertTrue(response.context["downstream_handoff_rows"])
        self.assertTrue(response.context["trunk_handoff_rows"])
        self.assertTrue(response.context["erp_critical_path_rows"])
        self.assertTrue(response.context["erp_release_rows"])
        self.assertTrue(response.context["erp_article_chain_rows"])

    def test_insumo_update_shows_recent_commercial_signal(self):
        receta = Receta.objects.create(
            nombre="Pastel Comercial Maestro",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Pastel",
            hash_contenido="hash-maestro-commercial-001",
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=self.insumo,
            insumo_texto=self.insumo.nombre,
            cantidad=Decimal("1"),
            unidad=self.unidad,
            unidad_texto="kgc",
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        sucursal = Sucursal.objects.create(codigo="001", nombre="Centro", activa=True)
        today = timezone.localdate()
        VentaHistorica.objects.create(
            receta=receta,
            sucursal=sucursal,
            fecha=today - timedelta(days=2),
            cantidad=Decimal("15"),
            tickets=4,
        )
        VentaHistorica.objects.create(
            receta=receta,
            sucursal=sucursal,
            fecha=today - timedelta(days=8),
            cantidad=Decimal("11"),
            tickets=3,
        )

        response = self.client.get(reverse("maestros:insumo_update", args=[self.insumo.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Impacto comercial reciente")
        self.assertContains(response, "Señal comercial")
        self.assertContains(response, "Pastel Comercial Maestro")
        self.assertIn("commercial_signal", response.context)
        self.assertEqual(response.context["commercial_signal"]["days_count"], 2)

    def test_insumo_list_shows_demand_priority_rows(self):
        receta = Receta.objects.create(
            nombre="Pastel Prioridad Maestro",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pastel",
            categoria="Pastel",
            hash_contenido="hash-maestro-priority-001",
        )
        insumo_critico = Insumo.objects.create(
            nombre="Caja Prioritaria Maestro",
            tipo_item=Insumo.TIPO_EMPAQUE,
            activo=True,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            insumo=insumo_critico,
            insumo_texto=insumo_critico.nombre,
            cantidad=Decimal("1"),
            unidad=self.unidad,
            unidad_texto="kgc",
            match_status=LineaReceta.STATUS_AUTO,
            match_score=100,
            match_method=LineaReceta.MATCH_EXACT,
        )
        sucursal = Sucursal.objects.create(codigo="009", nombre="Sucursal Prioridad", activa=True)
        today = timezone.localdate()
        VentaHistorica.objects.create(
            receta=receta,
            sucursal=sucursal,
            fecha=today - timedelta(days=3),
            cantidad=Decimal("82"),
            tickets=5,
        )

        response = self.client.get(reverse("maestros:insumo_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Artículos maestros prioritarios por demanda")
        self.assertContains(response, "Caja Prioritaria Maestro")
        self.assertContains(response, "Demanda crítica bloqueada")
        self.assertIn("demand_priority_rows", response.context)
        self.assertIn("critical_demand_priority_rows", response.context)
        self.assertIn("daily_critical_close_focus", response.context)
        self.assertTrue(response.context["demand_priority_rows"])
        self.assertTrue(response.context["critical_demand_priority_rows"])
        self.assertContains(response, "Cola crítica de cierre del maestro")
        self.assertContains(response, "Cierre prioritario del día")
        self.assertContains(response, "Liberación diaria retenida")
        self.assertEqual(response.context["demand_priority_summary"]["critical_count"], 1)
