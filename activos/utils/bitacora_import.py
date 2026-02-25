from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import StringIO
from pathlib import Path
from typing import Any, BinaryIO

from django.db import transaction

from openpyxl import load_workbook

from activos.models import Activo, BitacoraMantenimiento, OrdenMantenimiento


SECTION_HINTS = (
    "PRODUCCION",
    "VENTAS",
    "MATRIZ",
    "SUCURSAL",
    "ALMACEN",
    "OFICINA",
    "AREA",
)


@dataclass
class ParsedRow:
    nombre: str
    marca: str
    modelo: str
    serie: str
    ubicacion: str
    fecha_1: date | None
    costo_1: Decimal
    fecha_2: date | None
    costo_2: Decimal


def import_bitacora(
    archivo: str | Path | BinaryIO,
    *,
    sheet_name: str = "",
    dry_run: bool = False,
    skip_servicios: bool = False,
) -> dict[str, Any]:
    source = _build_source_rows(archivo=archivo, sheet_name=sheet_name)

    stats = {
        "sheet_name": source["sheet_name"],
        "source_format": source["source_format"],
        "filas_leidas": 0,
        "filas_validas": 0,
        "activos_creados": 0,
        "activos_actualizados": 0,
        "servicios_creados": 0,
        "servicios_omitidos": 0,
    }

    current_location = ""

    @transaction.atomic
    def _run():
        nonlocal current_location
        for row_idx, raw_row in source["rows"]:
            stats["filas_leidas"] += 1
            raw_name = _as_text(raw_row.get("nombre"))
            raw_brand = _as_text(raw_row.get("marca"))
            raw_model = _as_text(raw_row.get("modelo"))
            raw_serial = _as_text(raw_row.get("serie"))
            raw_date_1 = raw_row.get("fecha_1")
            raw_cost_1 = raw_row.get("costo_1")
            raw_date_2 = raw_row.get("fecha_2")
            raw_cost_2 = raw_row.get("costo_2")

            if not any(
                [raw_name, raw_brand, raw_model, raw_serial, raw_date_1, raw_cost_1, raw_date_2, raw_cost_2]
            ):
                continue

            if _is_section_row(raw_name, raw_brand, raw_model, raw_serial, raw_date_1, raw_cost_1, raw_date_2, raw_cost_2):
                current_location = raw_name.strip()
                continue

            if _is_header_row(raw_name, raw_brand, raw_model, raw_serial):
                continue

            if not raw_name:
                continue

            parsed = ParsedRow(
                nombre=raw_name.strip(),
                marca=raw_brand.strip(),
                modelo=raw_model.strip(),
                serie=raw_serial.strip(),
                ubicacion=current_location,
                fecha_1=_as_date(raw_date_1),
                costo_1=_as_decimal(raw_cost_1),
                fecha_2=_as_date(raw_date_2),
                costo_2=_as_decimal(raw_cost_2),
            )
            stats["filas_validas"] += 1

            activo, created, changed = _upsert_activo(parsed)
            if created:
                stats["activos_creados"] += 1
            elif changed:
                stats["activos_actualizados"] += 1

            if skip_servicios:
                continue

            for fecha, costo in ((parsed.fecha_1, parsed.costo_1), (parsed.fecha_2, parsed.costo_2)):
                if not fecha:
                    continue
                if _ensure_service_order(activo, fecha, costo):
                    stats["servicios_creados"] += 1
                else:
                    stats["servicios_omitidos"] += 1

        if dry_run:
            transaction.set_rollback(True)

    _run()
    return stats


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _build_source_rows(archivo: str | Path | BinaryIO, sheet_name: str) -> dict[str, Any]:
    source_name = _get_source_name(archivo)
    ext = Path(source_name).suffix.lower()

    if ext in {".csv", ".tsv"}:
        rows = _iter_csv_rows(archivo)
        return {
            "sheet_name": "CSV",
            "source_format": "CSV",
            "rows": rows,
        }

    wb = load_workbook(archivo, data_only=True)
    selected_sheet_name = sheet_name or wb.sheetnames[0]
    if selected_sheet_name not in wb.sheetnames:
        raise ValueError(f"La hoja '{selected_sheet_name}' no existe. Hojas: {', '.join(wb.sheetnames)}")
    ws = wb[selected_sheet_name]
    return {
        "sheet_name": selected_sheet_name,
        "source_format": "XLSX",
        "rows": _iter_xlsx_rows(ws),
    }


def _iter_xlsx_rows(ws) -> list[tuple[int, dict[str, Any]]]:
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_idx in range(1, ws.max_row + 1):
        rows.append(
            (
                row_idx,
                {
                    "nombre": ws.cell(row_idx, 2).value,
                    "marca": ws.cell(row_idx, 3).value,
                    "modelo": ws.cell(row_idx, 4).value,
                    "serie": ws.cell(row_idx, 5).value,
                    "fecha_1": ws.cell(row_idx, 6).value,
                    "costo_1": ws.cell(row_idx, 7).value,
                    "fecha_2": ws.cell(row_idx, 8).value,
                    "costo_2": ws.cell(row_idx, 9).value,
                },
            )
        )
    return rows


def _get_source_name(archivo: str | Path | BinaryIO) -> str:
    if isinstance(archivo, (str, Path)):
        return str(archivo)
    return str(getattr(archivo, "name", "bitacora.xlsx"))


def _decode_csv_bytes(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _read_csv_text(archivo: str | Path | BinaryIO) -> str:
    if isinstance(archivo, (str, Path)):
        raw = Path(archivo).read_bytes()
        return _decode_csv_bytes(raw)

    if hasattr(archivo, "seek"):
        archivo.seek(0)
    raw = archivo.read()
    if hasattr(archivo, "seek"):
        archivo.seek(0)
    if isinstance(raw, str):
        return raw
    return _decode_csv_bytes(raw)


def _iter_csv_rows(archivo: str | Path | BinaryIO) -> list[tuple[int, dict[str, Any]]]:
    text = _read_csv_text(archivo)
    sample = text[:2048]
    delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
    reader = csv.reader(StringIO(text), delimiter=delimiter)
    all_rows = list(reader)
    if not all_rows:
        return []

    header_map = _detect_csv_header(all_rows[0])
    payload_rows = all_rows[1:] if header_map else all_rows
    parsed_rows: list[tuple[int, dict[str, Any]]] = []
    start_idx = 2 if header_map else 1
    for idx, row in enumerate(payload_rows, start=start_idx):
        parsed_rows.append((idx, _extract_csv_row(row, header_map)))
    return parsed_rows


def _detect_csv_header(first_row: list[str]) -> dict[str, int]:
    normalized = [_normalize_header(col) for col in first_row]
    if not normalized:
        return {}

    mapping = {}
    for key, aliases in {
        "nombre": {"nombre", "equipo", "activo"},
        "marca": {"marca"},
        "modelo": {"modelo"},
        "serie": {"serie", "serie:"},
        "fecha_1": {"fecha_mantenimiento", "fecha_1", "fecha_mantenimiento_1", "fecha_servicio_1"},
        "costo_1": {"costo", "costo_1", "costo_servicio_1"},
        "fecha_2": {"fecha_2", "fecha_mantenimiento_2", "fecha_servicio_2"},
        "costo_2": {"costo_2", "costo_servicio_2"},
    }.items():
        for i, col in enumerate(normalized):
            if col in aliases:
                mapping[key] = i
                break

    # Consider header valid if at least core fields detected.
    if "nombre" in mapping and ("marca" in mapping or "modelo" in mapping or "serie" in mapping):
        return mapping
    return {}


def _extract_csv_row(row: list[str], header_map: dict[str, int]) -> dict[str, Any]:
    def _get(index: int) -> str:
        if index < 0 or index >= len(row):
            return ""
        return row[index]

    if header_map:
        return {
            "nombre": _get(header_map.get("nombre", -1)),
            "marca": _get(header_map.get("marca", -1)),
            "modelo": _get(header_map.get("modelo", -1)),
            "serie": _get(header_map.get("serie", -1)),
            "fecha_1": _get(header_map.get("fecha_1", -1)),
            "costo_1": _get(header_map.get("costo_1", -1)),
            "fecha_2": _get(header_map.get("fecha_2", -1)),
            "costo_2": _get(header_map.get("costo_2", -1)),
        }

    # Legacy positional format: B-I in spreadsheets, but allow compact 8-column CSV.
    name_idx = 1 if len(row) >= 9 else 0
    return {
        "nombre": _get(name_idx),
        "marca": _get(name_idx + 1),
        "modelo": _get(name_idx + 2),
        "serie": _get(name_idx + 3),
        "fecha_1": _get(name_idx + 4),
        "costo_1": _get(name_idx + 5),
        "fecha_2": _get(name_idx + 6),
        "costo_2": _get(name_idx + 7),
    }


def _as_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _as_decimal(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0")


def _as_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _is_header_row(name: str, brand: str, model: str, serial: str) -> bool:
    key = " ".join([name, brand, model, serial]).upper()
    return "FECHA MANTENIMIENTO" in key or ("MARCA" in key and "MODELO" in key)


def _is_section_row(name: str, brand: str, model: str, serial: str, d1, c1, d2, c2) -> bool:
    if not name:
        return False
    if any([brand, model, serial, d1, c1, d2, c2]):
        return False
    upper = name.upper()
    return any(h in upper for h in SECTION_HINTS)


def _infer_categoria(nombre: str) -> str:
    n = (nombre or "").upper()
    if "HORNO" in n:
        return "Hornos"
    if "AIRE" in n or "MINISPLIT" in n:
        return "Aire acondicionado"
    if "REFRIG" in n or "FREEZER" in n or "CUARTO FRIO" in n:
        return "Refrigeración"
    if "BATIDORA" in n:
        return "Batidoras"
    if "BASCULA" in n:
        return "Básculas"
    if "LICUADORA" in n:
        return "Licuadoras"
    return "Equipos"


def _compose_notes(marca: str, modelo: str, serie: str, previous: str) -> str:
    tags = []
    if marca:
        tags.append(f"Marca: {marca}")
    if modelo:
        tags.append(f"Modelo: {modelo}")
    if serie:
        tags.append(f"Serie: {serie}")
    base = " | ".join(tags)
    if previous and base and base not in previous:
        return (previous + "\n" + base).strip()
    if previous:
        return previous
    return base


def _upsert_activo(parsed: ParsedRow) -> tuple[Activo, bool, bool]:
    qs = Activo.objects.filter(nombre__iexact=parsed.nombre)
    if parsed.ubicacion:
        qs = qs.filter(ubicacion__iexact=parsed.ubicacion)
    activo = qs.order_by("id").first()

    created = False
    changed = False
    if not activo:
        activo = Activo(
            nombre=parsed.nombre,
            categoria=_infer_categoria(parsed.nombre),
            ubicacion=parsed.ubicacion,
            estado=Activo.ESTADO_OPERATIVO,
            criticidad=Activo.CRITICIDAD_MEDIA,
            activo=True,
        )
        created = True

    new_notes = _compose_notes(parsed.marca, parsed.modelo, parsed.serie, activo.notas)
    if parsed.ubicacion and activo.ubicacion != parsed.ubicacion:
        activo.ubicacion = parsed.ubicacion
        changed = True
    if new_notes != (activo.notas or ""):
        activo.notas = new_notes
        changed = True
    if not activo.categoria:
        activo.categoria = _infer_categoria(parsed.nombre)
        changed = True

    if created or changed:
        activo.save()

    return activo, created, changed


def _ensure_service_order(activo: Activo, fecha: date, costo: Decimal) -> bool:
    existing = OrdenMantenimiento.objects.filter(
        activo_ref=activo,
        tipo=OrdenMantenimiento.TIPO_PREVENTIVO,
        fecha_programada=fecha,
        estatus=OrdenMantenimiento.ESTATUS_CERRADA,
    ).first()
    if existing:
        return False

    orden = OrdenMantenimiento.objects.create(
        activo_ref=activo,
        tipo=OrdenMantenimiento.TIPO_PREVENTIVO,
        prioridad=OrdenMantenimiento.PRIORIDAD_MEDIA,
        estatus=OrdenMantenimiento.ESTATUS_CERRADA,
        fecha_programada=fecha,
        fecha_inicio=fecha,
        fecha_cierre=fecha,
        responsable="Servicio externo",
        descripcion="Servicio importado desde bitácora histórica",
        costo_otros=costo,
    )
    BitacoraMantenimiento.objects.create(
        orden=orden,
        accion="IMPORT_SERVICIO",
        comentario="Registro importado desde archivo histórico",
        usuario=None,
        costo_adicional=Decimal("0"),
    )
    return True
