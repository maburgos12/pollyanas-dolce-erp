import csv
from datetime import timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db.models import Count, Q, Sum
from django.db.models.functions import Lower
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook

from core.access import can_manage_inventario, can_view_inventario
from core.audit import log_event
from core.models import AuditLog
from maestros.models import Proveedor

from .models import Activo, BitacoraMantenimiento, OrdenMantenimiento, PlanMantenimiento
from .utils.bitacora_import import import_bitacora


def _safe_decimal(value) -> Decimal:
    try:
        return Decimal(str(value or 0))
    except Exception:
        return Decimal("0")


def _safe_int(value, default: int = 0) -> int:
    try:
        return int(str(value or "").strip())
    except Exception:
        return default


def _parse_date(value):
    raw = (value or "").strip()
    if not raw:
        return None
    try:
        return timezone.datetime.fromisoformat(raw).date()
    except Exception:
        return None


def _module_tabs(active: str) -> list[dict]:
    return [
        {"label": "Dashboard", "url_name": "activos:dashboard", "active": active == "dashboard"},
        {"label": "Activos", "url_name": "activos:activos", "active": active == "activos"},
        {"label": "Planes", "url_name": "activos:planes", "active": active == "planes"},
        {"label": "Órdenes", "url_name": "activos:ordenes", "active": active == "ordenes"},
        {"label": "Reportes", "url_name": "activos:reportes", "active": active == "reportes"},
        {"label": "Calendario", "url_name": "activos:calendario", "active": active == "calendario"},
    ]


def _activos_enterprise_chain(
    *,
    activos_total: int,
    blockers_total: int,
    planes_total: int,
    overdue_plans: int,
    ordenes_abiertas: int,
    critical_open: int,
    bitacora_30d: int,
) -> list[dict]:
    activos_listos = max(activos_total - blockers_total, 0)
    chain = [
        {
            "step": "01",
            "title": "Maestro de activos",
            "detail": "Catálogo, clasificación, proveedor y criticidad del equipo.",
            "count": activos_listos,
            "status": "Listos ERP" if blockers_total == 0 else f"{blockers_total} con brecha",
            "tone": "success" if blockers_total == 0 else "warning",
            "url": reverse("activos:activos"),
            "cta": "Abrir maestro",
            "owner": "Activos / Mantenimiento",
            "next_step": "Cerrar clasificación, proveedor y criticidad del activo.",
        },
        {
            "step": "02",
            "title": "Planes preventivos",
            "detail": "Cobertura activa, responsables y vencimientos programados.",
            "count": planes_total,
            "status": "Bajo control" if overdue_plans == 0 else f"{overdue_plans} vencidos",
            "tone": "success" if overdue_plans == 0 else "danger",
            "url": reverse("activos:planes"),
            "cta": "Abrir planes",
            "owner": "Planeación / Operación",
            "next_step": "Regularizar vencimientos y asegurar próxima ejecución.",
        },
        {
            "step": "03",
            "title": "Órdenes documentales",
            "detail": "Correctivos y preventivos con seguimiento operativo y costo.",
            "count": ordenes_abiertas,
            "status": "Estable" if critical_open == 0 else f"{critical_open} críticas",
            "tone": "success" if critical_open == 0 else "danger",
            "url": reverse("activos:ordenes"),
            "cta": "Abrir órdenes",
            "owner": "Mantenimiento / Ejecución",
            "next_step": "Cerrar correctivos críticos y documentar costo real.",
        },
        {
            "step": "04",
            "title": "Bitácora y trazabilidad",
            "detail": "Eventos recientes documentados para auditoría y cierre.",
            "count": bitacora_30d,
            "status": "Con trazabilidad" if bitacora_30d else "Sin eventos recientes",
            "tone": "success" if bitacora_30d else "warning",
            "url": reverse("activos:reportes"),
            "cta": "Abrir bitácora",
            "owner": "DG / Auditoría",
            "next_step": "Conservar evidencia cerrada y trazabilidad reciente.",
        },
    ]
    for index, item in enumerate(chain):
        previous = chain[index - 1] if index else None
        item["depends_on"] = previous["title"] if previous else "Origen del módulo"
        item["completion"] = 100 if item["tone"] == "success" else (60 if item["tone"] == "warning" else 25)
        if previous:
            previous_blocking = previous.get("tone") != "success"
            item["dependency_status"] = (
                f"Condicionado por {previous['title'].lower()}"
                if previous_blocking
                else f"Listo desde {previous['title'].lower()}"
            )
        else:
            item["dependency_status"] = "Punto de arranque del módulo"
    return chain

def _activos_maturity_summary(*, chain: list[dict]) -> dict:
    total_steps = len(chain)
    completed_steps = sum(1 for item in chain if item.get("tone") == "success")
    attention_steps = max(total_steps - completed_steps, 0)
    coverage_pct = int(round((completed_steps / total_steps) * 100)) if total_steps else 0
    next_priority = next((item for item in chain if item.get("tone") != "success"), None)
    return {
        "completed_steps": completed_steps,
        "attention_steps": attention_steps,
        "coverage_pct": coverage_pct,
        "next_priority_title": next_priority.get("title", "Cadena estabilizada") if next_priority else "Cadena estabilizada",
        "next_priority_detail": next_priority.get("detail", "Sin bloqueos abiertos en activos.") if next_priority else "Sin bloqueos abiertos en activos.",
        "next_priority_url": next_priority.get("url", reverse("activos:dashboard")) if next_priority else reverse("activos:dashboard"),
        "next_priority_cta": next_priority.get("cta", "Abrir dashboard") if next_priority else "Abrir dashboard",
    }


def _activos_handoff_map(*, blockers_total: int, overdue_plans: int, ordenes_abiertas: int, bitacora_30d: int) -> list[dict]:
    handoffs = [
        {
            "label": "Maestro -> Planes",
            "detail": "Un activo incompleto o sin plan impide la cobertura preventiva.",
            "count": blockers_total,
            "tone": "success" if blockers_total == 0 else "warning",
            "status": "Controlado" if blockers_total == 0 else "Con brechas",
            "url": reverse("activos:activos"),
            "cta": "Abrir activos",
            "owner": "Activos / Mantenimiento",
            "depends_on": "Alta del activo + criticidad",
            "exit_criteria": "El activo ya quedó clasificado y listo para sostener un plan preventivo vigente.",
            "next_step": "Cerrar brechas del maestro del activo.",
            "completion": 100 if blockers_total == 0 else 55,
        },
        {
            "label": "Planes -> Órdenes",
            "detail": "Los planes vencidos deben convertirse o regularizarse como ejecución documentada.",
            "count": overdue_plans,
            "tone": "success" if overdue_plans == 0 else "danger",
            "status": "Al día" if overdue_plans == 0 else "Con atraso",
            "url": reverse("activos:planes"),
            "cta": "Abrir planes",
            "owner": "Planeación / Operación",
            "depends_on": "Cobertura preventiva vigente",
            "exit_criteria": "Los planes ya quedaron activos y sin vencimientos que frenen la ejecución documental.",
            "next_step": "Regularizar planes vencidos y programar ejecución.",
            "completion": 100 if overdue_plans == 0 else 25,
        },
        {
            "label": "Órdenes -> Bitácora",
            "detail": "Toda intervención debe dejar orden cerrada y trazabilidad reciente.",
            "count": ordenes_abiertas,
            "tone": "success" if ordenes_abiertas == 0 and bitacora_30d > 0 else "warning",
            "status": "Controlado" if ordenes_abiertas == 0 and bitacora_30d > 0 else "Seguimiento abierto",
            "url": reverse("activos:ordenes"),
            "cta": "Abrir órdenes",
            "owner": "Mantenimiento / Ejecución",
            "depends_on": "Órdenes con costo + evidencia",
            "exit_criteria": "Las órdenes ya cerraron y la bitácora reciente deja trazabilidad auditable.",
            "next_step": "Cerrar órdenes abiertas y documentar bitácora.",
            "completion": 100 if ordenes_abiertas == 0 and bitacora_30d > 0 else 60,
        },
    ]
    return handoffs


def _activos_operational_health_cards(*, cards: list[dict], module_key: str) -> list[dict]:
    label_map = {
        "dashboard": {
            "warning": "Activos por regularizar",
            "danger": "Bloqueos críticos",
        },
        "activos": {
            "warning": "Gobierno del maestro",
            "danger": "Bloqueos críticos",
        },
        "calendario": {
            "warning": "Eventos por atender",
            "danger": "Riesgo en curso",
        },
        "planes": {
            "warning": "Planes por regularizar",
            "danger": "Riesgo preventivo",
        },
        "ordenes": {
            "warning": "Órdenes por completar",
            "danger": "Riesgo correctivo",
        },
        "reportes": {
            "warning": "Reportes abiertos",
            "danger": "Atención urgente",
        },
    }
    module_labels = label_map.get(module_key, {})
    warning_total = sum(int(card.get("count") or 0) for card in cards if card.get("tone") == "warning")
    danger_total = sum(int(card.get("count") or 0) for card in cards if card.get("tone") == "danger")
    return [
        {
            "label": module_labels.get("warning", "Pendientes operativas"),
            "count": warning_total,
            "tone": "warning",
        },
        {
            "label": module_labels.get("danger", "Bloqueos críticos"),
            "count": danger_total,
            "tone": "danger",
        },
    ]


def _activos_document_owner(label: str) -> str:
    normalized = (label or "").strip().lower()
    if "activo" in normalized:
        return "Activos / Mantenimiento"
    if "plan" in normalized:
        return "Planeación / Operación"
    if "orden" in normalized:
        return "Mantenimiento / Ejecución"
    if "bitácora" in normalized or "bitacora" in normalized:
        return "DG / Auditoría"
    if "reporte" in normalized:
        return "Mantenimiento / Ejecución"
    return "Operación ERP"


def _activos_command_center(
    *,
    owner: str,
    blockers_total: int,
    maturity_summary: dict,
    default_url: str,
    default_cta: str,
) -> dict:
    attention_steps = int(maturity_summary.get("attention_steps") or 0)
    next_priority_title = maturity_summary.get("next_priority_title") or "Cadena operativa estabilizada"
    next_priority_detail = maturity_summary.get("next_priority_detail") or "Sin acciones pendientes."
    next_priority_url = maturity_summary.get("next_priority_url") or default_url
    next_priority_cta = maturity_summary.get("next_priority_cta") or default_cta
    if blockers_total > 0:
        status = "Con bloqueos"
        tone = "danger"
    elif attention_steps > 0:
        status = "En seguimiento"
        tone = "warning"
    else:
        status = "Estable"
        tone = "success"
    return {
        "owner": owner,
        "status": status,
        "tone": tone,
        "blockers": blockers_total,
        "next_step": f"{next_priority_title}. {next_priority_detail}",
        "url": next_priority_url,
        "cta": next_priority_cta,
    }


def _annotate_activos_document_stage_rows(rows: list[dict]) -> list[dict]:
    annotated_rows: list[dict] = []
    for row in rows:
        normalized = dict(row)
        normalized["owner"] = normalized.get("owner") or _activos_document_owner(str(normalized.get("label", "")))
        total = int(normalized.get("open", 0) or 0) + int(normalized.get("closed", 0) or 0)
        normalized["completion"] = normalized.get("completion")
        if normalized["completion"] is None:
            normalized["completion"] = round((int(normalized.get("closed", 0) or 0) / max(total, 1)) * 100) if total else 0
        normalized["next_step"] = normalized.get("next_step") or "Abrir etapa"
        annotated_rows.append(normalized)
    return annotated_rows


def _activos_erp_governance_rows(rows: list[dict]) -> list[dict]:
    governance_rows: list[dict] = []
    for row in rows:
        governance_rows.append(
            {
                "front": row.get("label", "Frente operativo"),
                "owner": row.get("owner") or "Operación ERP",
                "blockers": int(row.get("open", 0) or 0),
                "completion": int(row.get("completion", 0) or 0),
                "detail": row.get("detail", ""),
                "next_step": row.get("next_step") or "Revisar etapa",
                "url": row.get("url") or reverse("activos:dashboard"),
                "cta": "Abrir",
            }
        )
    return governance_rows


def _activos_release_gate_rows(rows: list[dict]) -> list[dict]:
    release_rows: list[dict] = []
    for index, row in enumerate(rows, start=1):
        open_count = int(row.get("open", 0) or 0)
        closed = int(row.get("closed", 0) or 0)
        total = max(open_count + closed, 1)
        release_rows.append(
            {
                "step": f"{index:02d}",
                "title": row.get("label", f"Etapa {index}"),
                "detail": row.get("detail", ""),
                "completed": closed,
                "open_count": open_count,
                "total": total,
                "tone": "success" if open_count == 0 else "warning",
                "url": row.get("url") or reverse("activos:dashboard"),
                "cta": row.get("cta") or "Abrir",
            }
        )
    return release_rows


def _activos_release_gate_completion(rows: list[dict]) -> int:
    if not rows:
        return 0
    total = sum(int(row.get("total", 0) or 0) for row in rows)
    if total <= 0:
        return 0
    completed = sum(int(row.get("completed", 0) or 0) for row in rows)
    return round((completed / total) * 100)


def _activos_critical_path_rows(chain: list[dict[str, object]]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for index, item in enumerate(chain, start=1):
        tone = item.get("tone") or "warning"
        rows.append(
            {
                "priority": f"{index:02d}",
                "title": item.get("title", f"Tramo {index}"),
                "owner": item.get("owner", "Activos / Mantenimiento"),
                "blockers": 0 if tone == "success" else int(item.get("count") or 0),
                "completion": int(item.get("completion") or 0),
                "depends_on": item.get("depends_on", "Origen del módulo"),
                "detail": item.get("detail", ""),
                "next_step": item.get("next_step", "Revisar tramo"),
                "url": item.get("url", reverse("activos:dashboard")),
                "cta": item.get("cta", "Abrir"),
            }
        )
    return rows


def _activos_executive_radar_rows(
    document_stage_rows: list[dict[str, object]],
    enterprise_chain: list[dict[str, object]],
) -> list[dict[str, object]]:
    chain_by_step = {
        str(item.get("step") or ""): item
        for item in enterprise_chain
    }
    rows: list[dict[str, object]] = []
    for index, stage in enumerate(document_stage_rows, start=1):
        chain = chain_by_step.get(f"{index:02d}") or enterprise_chain[min(index - 1, len(enterprise_chain) - 1)] if enterprise_chain else {}
        open_count = int(stage.get("open", 0) or 0)
        completion = int(stage.get("completion", 0) or 0)
        if open_count <= 0 and completion >= 100:
            tone = "success"
            dominant_blocker = "Sin bloqueo activo"
        elif open_count > 0 and completion < 60:
            tone = "danger"
            dominant_blocker = stage.get("detail") or "Bloqueo operativo abierto"
        else:
            tone = "warning"
            dominant_blocker = stage.get("detail") or "Seguimiento operativo pendiente"
        rows.append(
            {
                "phase": stage.get("label", f"Fase {index}"),
                "owner": stage.get("owner") or "Activos / Mantenimiento",
                "status": "Controlado" if tone == "success" else "En seguimiento" if tone == "warning" else "Con bloqueo",
                "tone": tone,
                "blockers": open_count,
                "progress_pct": completion,
                "dominant_blocker": dominant_blocker,
                "depends_on": chain.get("depends_on", "Origen del módulo"),
                "dependency_status": chain.get("dependency_status", "Sin dependencia registrada"),
                "next_step": stage.get("next_step") or chain.get("next_step") or "Abrir fase",
                "url": stage.get("url") or chain.get("url") or reverse("activos:dashboard"),
                "cta": chain.get("cta") or "Abrir",
            }
        )
    return rows


def _activos_focus_cards(*, selected_focus: str, enterprise_cards: list[dict], module_key: str) -> list[dict]:
    focus_map = {
        "activos": {
            "SIN_CATEGORIA": "Completar clasificación del activo para gobierno y reportes.",
            "SIN_PROVEEDOR": "Asignar proveedor principal de mantenimiento por equipo.",
            "SIN_PLAN": "Crear cobertura preventiva para no dejar activos fuera del plan.",
            "PLAN_VENCIDO": "Regularizar vencimientos antes de generar riesgo operativo.",
            "ORDEN_CRITICA": "Cerrar correctivos críticos abiertos en el catálogo.",
            "FUERA_SERVICIO_CRITICO": "Recuperar activos críticos fuera de servicio o justificar reemplazo.",
        }
    }
    module_focus = focus_map.get(module_key, {})
    base_url = reverse("activos:activos")
    cards: list[dict] = []
    for item in enterprise_cards:
        cards.append(
            {
                "key": item["key"],
                "label": item["label"],
                "count": item["count"],
                "tone": item["tone"],
                "detail": module_focus.get(item["key"], "Vista enfocada para cierre documental."),
                "active": selected_focus == item["key"],
                "url": f"{base_url}?master_gap={item['key']}",
            }
        )
    return cards


def _active_focus_summary(*, focus_label: str, focus_detail: str, clear_url: str) -> dict:
    return {
        "label": focus_label,
        "detail": focus_detail,
        "clear_url": clear_url,
    }


def _prioridad_por_criticidad(criticidad: str) -> str:
    if criticidad == Activo.CRITICIDAD_ALTA:
        return OrdenMantenimiento.PRIORIDAD_ALTA
    if criticidad == Activo.CRITICIDAD_BAJA:
        return OrdenMantenimiento.PRIORIDAD_BAJA
    return OrdenMantenimiento.PRIORIDAD_MEDIA


def _activo_enterprise_profile(
    activo: Activo,
    *,
    overdue_plan_asset_ids: set[int],
    active_plan_asset_ids: set[int],
    critical_open_asset_ids: set[int],
) -> dict:
    gaps: list[dict[str, str]] = []
    if not (activo.categoria or "").strip():
        gaps.append(
            {
                "key": "SIN_CATEGORIA",
                "label": "Sin categoría",
                "detail": "Clasificar el activo para reportes y mantenimiento.",
            }
        )
    if activo.proveedor_mantenimiento_id is None:
        gaps.append(
            {
                "key": "SIN_PROVEEDOR",
                "label": "Sin proveedor",
                "detail": "Asignar proveedor principal de mantenimiento.",
            }
        )
    if activo.id not in active_plan_asset_ids:
        gaps.append(
            {
                "key": "SIN_PLAN",
                "label": "Sin plan",
                "detail": "Crear plan preventivo/calibración para el activo.",
            }
        )
    if activo.id in overdue_plan_asset_ids:
        gaps.append(
            {
                "key": "PLAN_VENCIDO",
                "label": "Plan vencido",
                "detail": "Regularizar mantenimiento programado vencido.",
            }
        )
    if activo.id in critical_open_asset_ids:
        gaps.append(
            {
                "key": "ORDEN_CRITICA",
                "label": "Orden crítica abierta",
                "detail": "Cerrar orden crítica o mover a en proceso con responsable.",
            }
        )
    if activo.estado == Activo.ESTADO_FUERA_SERVICIO and activo.criticidad == Activo.CRITICIDAD_ALTA:
        gaps.append(
            {
                "key": "FUERA_SERVICIO_CRITICO",
                "label": "Fuera de servicio crítico",
                "detail": "Activo crítico fuera de servicio; priorizar recuperación o reemplazo.",
            }
        )

    if any(g["key"] in {"FUERA_SERVICIO_CRITICO", "ORDEN_CRITICA", "PLAN_VENCIDO"} for g in gaps):
        status_label = "Crítico"
        status_tone = "danger"
    elif gaps:
        status_label = "Pendiente"
        status_tone = "warning"
    elif activo.estado == Activo.ESTADO_MANTENIMIENTO:
        status_label = "Seguimiento"
        status_tone = "warning"
    else:
        status_label = "Listo ERP"
        status_tone = "success"

    next_action = "Operación estable"
    if gaps:
        next_action = gaps[0]["detail"]
    elif activo.estado == Activo.ESTADO_MANTENIMIENTO:
        next_action = "Cerrar orden o devolver a operativo."

    return {
        "gaps": gaps,
        "status_label": status_label,
        "status_tone": status_tone,
        "next_action": next_action,
        "blocks_operations": bool(gaps),
    }


def _plan_enterprise_profile(plan: PlanMantenimiento, *, today) -> dict:
    gaps: list[dict[str, str]] = []
    if not (plan.responsable or "").strip():
        gaps.append(
            {
                "key": "SIN_RESPONSABLE",
                "label": "Sin responsable",
                "detail": "Asignar responsable del plan.",
            }
        )
    if plan.proxima_ejecucion is None:
        gaps.append(
            {
                "key": "SIN_PROXIMA",
                "label": "Sin próxima ejecución",
                "detail": "Definir próxima ejecución del plan.",
            }
        )
    if (
        plan.estatus == PlanMantenimiento.ESTATUS_ACTIVO
        and plan.proxima_ejecucion
        and plan.proxima_ejecucion < today
    ):
        gaps.append(
            {
                "key": "VENCIDO",
                "label": "Vencido",
                "detail": "Registrar ejecución o generar orden preventiva.",
            }
        )
    if plan.activo_ref and plan.activo_ref.estado == Activo.ESTADO_FUERA_SERVICIO:
        gaps.append(
            {
                "key": "ACTIVO_FUERA_SERVICIO",
                "label": "Activo fuera de servicio",
                "detail": "Revisar vigencia del plan o recuperar el activo.",
            }
        )

    if any(g["key"] in {"VENCIDO", "ACTIVO_FUERA_SERVICIO"} for g in gaps):
        status_label = "Crítico"
        status_tone = "danger"
    elif gaps:
        status_label = "Pendiente"
        status_tone = "warning"
    else:
        status_label = "Listo ERP"
        status_tone = "success"
    next_action = gaps[0]["detail"] if gaps else "Plan controlado"
    return {"gaps": gaps, "status_label": status_label, "status_tone": status_tone, "next_action": next_action}


def _orden_enterprise_profile(orden: OrdenMantenimiento) -> dict:
    gaps: list[dict[str, str]] = []
    if not (orden.responsable or "").strip():
        gaps.append(
            {
                "key": "SIN_RESPONSABLE",
                "label": "Sin responsable",
                "detail": "Asignar responsable de atención/cierre.",
            }
        )
    if orden.estatus == OrdenMantenimiento.ESTATUS_EN_PROCESO and not orden.fecha_inicio:
        gaps.append(
            {
                "key": "SIN_FECHA_INICIO",
                "label": "Sin fecha inicio",
                "detail": "Registrar fecha de arranque real.",
            }
        )
    if orden.estatus == OrdenMantenimiento.ESTATUS_CERRADA and orden.costo_total <= Decimal("0"):
        gaps.append(
            {
                "key": "SIN_COSTO_CIERRE",
                "label": "Cierre sin costo",
                "detail": "Registrar costos del servicio cerrado.",
            }
        )
    if orden.tipo == OrdenMantenimiento.TIPO_PREVENTIVO and orden.plan_ref_id is None:
        gaps.append(
            {
                "key": "SIN_PLAN_ORIGEN",
                "label": "Sin plan origen",
                "detail": "Vincular orden preventiva a un plan.",
            }
        )
    if orden.prioridad in {OrdenMantenimiento.PRIORIDAD_CRITICA, OrdenMantenimiento.PRIORIDAD_ALTA} and orden.estatus in {
        OrdenMantenimiento.ESTATUS_PENDIENTE,
        OrdenMantenimiento.ESTATUS_EN_PROCESO,
    }:
        gaps.append(
            {
                "key": "ABIERTA_CRITICA",
                "label": "Abierta crítica",
                "detail": "Cerrar o estabilizar orden crítica.",
            }
        )

    if any(g["key"] in {"ABIERTA_CRITICA", "SIN_COSTO_CIERRE"} for g in gaps):
        status_label = "Crítico"
        status_tone = "danger"
    elif gaps:
        status_label = "Pendiente"
        status_tone = "warning"
    else:
        status_label = "Listo ERP"
        status_tone = "success"
    next_action = gaps[0]["detail"] if gaps else "Documento controlado"
    return {"gaps": gaps, "status_label": status_label, "status_tone": status_tone, "next_action": next_action}


def _build_activos_depuracion_rows(activos_rows: list[Activo], *, all_name_counts: dict[str, int]) -> list[dict]:
    suspect_exact = {
        "matriz",
        "crucero",
        "colosio",
        "payan",
        "n i o",
        "nio",
        "tunel",
        "leyva",
        "logistica",
    }
    dep_rows = []
    for activo in activos_rows:
        nombre = (activo.nombre or "").strip()
        nombre_norm = nombre.lower()
        notas = (activo.notas or "").strip()

        motivos = []
        acciones = []
        if not notas:
            motivos.append("Sin detalle técnico (marca/modelo/serie)")
            acciones.append("Completar notas con marca, modelo y serie")
        if all_name_counts.get(nombre_norm, 0) > 1:
            motivos.append("Nombre duplicado entre activos")
            acciones.append("Estandarizar nombre con sufijo de ubicación o código interno")
        if nombre_norm in suspect_exact:
            motivos.append("Nombre parece ubicación/departamento, no equipo")
            acciones.append("Renombrar con nombre real del equipo")
        if (activo.categoria or "").strip().lower() == "equipos":
            motivos.append("Categoría genérica")
            acciones.append("Reclasificar categoría operativa")
        if len(nombre.split()) <= 1 and len(nombre) <= 8:
            motivos.append("Nombre muy corto o ambiguo")
            acciones.append("Usar nombre descriptivo del activo")

        if motivos:
            dep_rows.append(
                {
                    "codigo": activo.codigo or "",
                    "nombre": nombre,
                    "ubicacion": (activo.ubicacion or "").strip(),
                    "categoria": (activo.categoria or "").strip(),
                    "estado": activo.estado,
                    "notas": notas,
                    "motivos": " | ".join(dict.fromkeys(motivos)),
                    "acciones_sugeridas": " | ".join(dict.fromkeys(acciones)),
                }
            )
    return dep_rows


def _export_activos_depuracion_csv(rows: list[dict]) -> HttpResponse:
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="activos_pendientes_depuracion_{timestamp}.csv"'
    writer = csv.writer(response)
    headers = ["codigo", "nombre", "ubicacion", "categoria", "estado", "notas", "motivos", "acciones_sugeridas"]
    writer.writerow(headers)
    for row in rows:
        writer.writerow([row.get(h, "") for h in headers])
    return response


def _export_bitacora_template_csv() -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="plantilla_bitacora_activos.csv"'
    writer = csv.writer(response)
    writer.writerow(["nombre", "marca", "modelo", "serie", "fecha_1", "costo_1", "fecha_2", "costo_2"])
    writer.writerow(["HORNO PRINCIPAL", "ACME", "HX-20", "SER-001", "2026-02-01", "1250.00", "", ""])
    return response


def _export_bitacora_template_xlsx() -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "bitacora"
    ws.append(["nombre", "marca", "modelo", "serie", "fecha_1", "costo_1", "fecha_2", "costo_2"])
    ws.append(["HORNO PRINCIPAL", "ACME", "HX-20", "SER-001", "2026-02-01", 1250.00, "", ""])
    for col, width in {"A": 28, "B": 18, "C": 18, "D": 18, "E": 14, "F": 12, "G": 14, "H": 12}.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="plantilla_bitacora_activos.xlsx"'
    return response


def _export_bitacora_runs_csv(runs) -> HttpResponse:
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="activos_import_bitacora_historial_{timestamp}.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "fecha",
            "usuario",
            "archivo",
            "modo",
            "formato",
            "hoja",
            "filas_leidas",
            "filas_validas",
            "activos_creados",
            "activos_actualizados",
            "servicios_creados",
            "servicios_omitidos",
        ]
    )
    for run in runs:
        payload = run.payload or {}
        writer.writerow(
            [
                timezone.localtime(run.timestamp).strftime("%Y-%m-%d %H:%M"),
                run.user.username if run.user else "Sistema",
                payload.get("filename", ""),
                "SIMULACION" if payload.get("dry_run") else "APLICADO",
                payload.get("source_format", ""),
                payload.get("sheet_name", ""),
                payload.get("filas_leidas", 0),
                payload.get("filas_validas", 0),
                payload.get("activos_creados", 0),
                payload.get("activos_actualizados", 0),
                payload.get("servicios_creados", 0),
                payload.get("servicios_omitidos", 0),
            ]
        )
    return response


def _export_bitacora_runs_xlsx(runs) -> HttpResponse:
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    wb = Workbook()
    ws = wb.active
    ws.title = "historial_importaciones"
    ws.append(
        [
            "fecha",
            "usuario",
            "archivo",
            "modo",
            "formato",
            "hoja",
            "filas_leidas",
            "filas_validas",
            "activos_creados",
            "activos_actualizados",
            "servicios_creados",
            "servicios_omitidos",
        ]
    )
    for run in runs:
        payload = run.payload or {}
        ws.append(
            [
                timezone.localtime(run.timestamp).strftime("%Y-%m-%d %H:%M"),
                run.user.username if run.user else "Sistema",
                payload.get("filename", ""),
                "SIMULACION" if payload.get("dry_run") else "APLICADO",
                payload.get("source_format", ""),
                payload.get("sheet_name", ""),
                payload.get("filas_leidas", 0),
                payload.get("filas_validas", 0),
                payload.get("activos_creados", 0),
                payload.get("activos_actualizados", 0),
                payload.get("servicios_creados", 0),
                payload.get("servicios_omitidos", 0),
            ]
        )
    for col, width in {
        "A": 18,
        "B": 16,
        "C": 34,
        "D": 12,
        "E": 12,
        "F": 14,
        "G": 12,
        "H": 12,
        "I": 14,
        "J": 18,
        "K": 16,
        "L": 16,
    }.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="activos_import_bitacora_historial_{timestamp}.xlsx"'
    return response


def _export_activos_depuracion_xlsx(rows: list[dict]) -> HttpResponse:
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    wb = Workbook()
    ws = wb.active
    ws.title = "pendientes_depuracion"
    headers = ["codigo", "nombre", "ubicacion", "categoria", "estado", "notas", "motivos", "acciones_sugeridas"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 48
    ws.column_dimensions["G"].width = 48
    ws.column_dimensions["H"].width = 48

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="activos_pendientes_depuracion_{timestamp}.xlsx"'
    return response


def _export_planes_csv(planes_rows: list[PlanMantenimiento]) -> HttpResponse:
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="activos_planes_{timestamp}.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "activo_codigo",
            "activo",
            "plan",
            "tipo",
            "estatus",
            "frecuencia_dias",
            "tolerancia_dias",
            "ultima_ejecucion",
            "proxima_ejecucion",
            "responsable",
        ]
    )
    for plan in planes_rows:
        writer.writerow(
            [
                plan.activo_ref.codigo if plan.activo_ref_id else "",
                plan.activo_ref.nombre if plan.activo_ref_id else "",
                plan.nombre,
                plan.tipo,
                plan.estatus,
                plan.frecuencia_dias,
                plan.tolerancia_dias,
                str(plan.ultima_ejecucion or ""),
                str(plan.proxima_ejecucion or ""),
                plan.responsable or "",
            ]
        )
    return response


def _export_planes_xlsx(planes_rows: list[PlanMantenimiento]) -> HttpResponse:
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    wb = Workbook()
    ws = wb.active
    ws.title = "planes"
    ws.append(
        [
            "activo_codigo",
            "activo",
            "plan",
            "tipo",
            "estatus",
            "frecuencia_dias",
            "tolerancia_dias",
            "ultima_ejecucion",
            "proxima_ejecucion",
            "responsable",
        ]
    )
    for plan in planes_rows:
        ws.append(
            [
                plan.activo_ref.codigo if plan.activo_ref_id else "",
                plan.activo_ref.nombre if plan.activo_ref_id else "",
                plan.nombre,
                plan.tipo,
                plan.estatus,
                plan.frecuencia_dias,
                plan.tolerancia_dias,
                str(plan.ultima_ejecucion or ""),
                str(plan.proxima_ejecucion or ""),
                plan.responsable or "",
            ]
        )
    for col, width in {
        "A": 16,
        "B": 32,
        "C": 32,
        "D": 14,
        "E": 14,
        "F": 16,
        "G": 16,
        "H": 18,
        "I": 18,
        "J": 24,
    }.items():
        ws.column_dimensions[col].width = width
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="activos_planes_{timestamp}.xlsx"'
    return response


def _export_ordenes_csv(ordenes_rows: list[OrdenMantenimiento]) -> HttpResponse:
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="activos_ordenes_{timestamp}.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "folio",
            "activo_codigo",
            "activo",
            "plan",
            "tipo",
            "prioridad",
            "estatus",
            "fecha_programada",
            "fecha_inicio",
            "fecha_cierre",
            "responsable",
            "costo_total",
            "descripcion",
        ]
    )
    for orden in ordenes_rows:
        writer.writerow(
            [
                orden.folio,
                orden.activo_ref.codigo if orden.activo_ref_id else "",
                orden.activo_ref.nombre if orden.activo_ref_id else "",
                orden.plan_ref.nombre if orden.plan_ref_id else "",
                orden.tipo,
                orden.prioridad,
                orden.estatus,
                str(orden.fecha_programada or ""),
                str(orden.fecha_inicio or ""),
                str(orden.fecha_cierre or ""),
                orden.responsable or "",
                str(orden.costo_total),
                orden.descripcion or "",
            ]
        )
    return response


def _export_ordenes_xlsx(ordenes_rows: list[OrdenMantenimiento]) -> HttpResponse:
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    wb = Workbook()
    ws = wb.active
    ws.title = "ordenes"
    ws.append(
        [
            "folio",
            "activo_codigo",
            "activo",
            "plan",
            "tipo",
            "prioridad",
            "estatus",
            "fecha_programada",
            "fecha_inicio",
            "fecha_cierre",
            "responsable",
            "costo_total",
            "descripcion",
        ]
    )
    for orden in ordenes_rows:
        ws.append(
            [
                orden.folio,
                orden.activo_ref.codigo if orden.activo_ref_id else "",
                orden.activo_ref.nombre if orden.activo_ref_id else "",
                orden.plan_ref.nombre if orden.plan_ref_id else "",
                orden.tipo,
                orden.prioridad,
                orden.estatus,
                str(orden.fecha_programada or ""),
                str(orden.fecha_inicio or ""),
                str(orden.fecha_cierre or ""),
                orden.responsable or "",
                float(orden.costo_total or 0),
                orden.descripcion or "",
            ]
        )
    for col, width in {
        "A": 18,
        "B": 16,
        "C": 28,
        "D": 28,
        "E": 14,
        "F": 12,
        "G": 14,
        "H": 18,
        "I": 14,
        "J": 14,
        "K": 24,
        "L": 14,
        "M": 56,
    }.items():
        ws.column_dimensions[col].width = width
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="activos_ordenes_{timestamp}.xlsx"'
    return response


def _export_reportes_servicio_csv(rows: list[dict]) -> HttpResponse:
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="activos_reportes_servicio_{timestamp}.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "folio",
            "fecha",
            "activo_codigo",
            "activo",
            "prioridad",
            "estatus",
            "semaforo",
            "dias",
            "descripcion",
            "responsable",
        ]
    )
    for item in rows:
        orden = item["orden"]
        writer.writerow(
            [
                orden.folio,
                str(orden.fecha_programada or ""),
                orden.activo_ref.codigo if orden.activo_ref_id else "",
                orden.activo_ref.nombre if orden.activo_ref_id else "",
                orden.prioridad,
                orden.estatus,
                item.get("semaforo_label", ""),
                item.get("dias", 0),
                orden.descripcion or "",
                orden.responsable or "",
            ]
        )
    return response


def _export_reportes_servicio_xlsx(rows: list[dict]) -> HttpResponse:
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M")
    wb = Workbook()
    ws = wb.active
    ws.title = "reportes_servicio"
    ws.append(
        [
            "folio",
            "fecha",
            "activo_codigo",
            "activo",
            "prioridad",
            "estatus",
            "semaforo",
            "dias",
            "descripcion",
            "responsable",
        ]
    )
    for item in rows:
        orden = item["orden"]
        ws.append(
            [
                orden.folio,
                str(orden.fecha_programada or ""),
                orden.activo_ref.codigo if orden.activo_ref_id else "",
                orden.activo_ref.nombre if orden.activo_ref_id else "",
                orden.prioridad,
                orden.estatus,
                item.get("semaforo_label", ""),
                int(item.get("dias", 0) or 0),
                orden.descripcion or "",
                orden.responsable or "",
            ]
        )
    for col, width in {
        "A": 18,
        "B": 14,
        "C": 16,
        "D": 32,
        "E": 12,
        "F": 14,
        "G": 12,
        "H": 10,
        "I": 56,
        "J": 24,
    }.items():
        ws.column_dimensions[col].width = width
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="activos_reportes_servicio_{timestamp}.xlsx"'
    return response


@login_required
def dashboard(request):
    if not can_view_inventario(request.user):
        raise PermissionDenied("No tienes permisos para ver Activos.")

    today = timezone.localdate()
    week_limit = today + timedelta(days=7)
    month_limit = today + timedelta(days=30)

    activos_qs = Activo.objects.filter(activo=True)
    ordenes_abiertas_qs = OrdenMantenimiento.objects.filter(
        estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO]
    )
    planes_activos_qs = PlanMantenimiento.objects.filter(estatus=PlanMantenimiento.ESTATUS_ACTIVO, activo=True)

    proximos = list(
        planes_activos_qs.filter(proxima_ejecucion__isnull=False, proxima_ejecucion__lte=month_limit)
        .select_related("activo_ref")
        .order_by("proxima_ejecucion", "id")[:30]
    )
    ordenes_recientes = list(
        OrdenMantenimiento.objects.select_related("activo_ref", "plan_ref")
        .order_by("-fecha_programada", "-id")[:20]
    )
    planes_vencidos_qs = list(
        planes_activos_qs.filter(proxima_ejecucion__isnull=False, proxima_ejecucion__lt=today)
        .select_related("activo_ref")
        .order_by("proxima_ejecucion", "id")[:20]
    )
    planes_vencidos_rows = [
        {
            "plan": plan,
            "dias_vencido": max((today - plan.proxima_ejecucion).days, 0) if plan.proxima_ejecucion else 0,
        }
        for plan in planes_vencidos_qs
    ]
    ordenes_criticas_rows = list(
        ordenes_abiertas_qs.filter(
            tipo=OrdenMantenimiento.TIPO_CORRECTIVO,
            prioridad__in=[OrdenMantenimiento.PRIORIDAD_CRITICA, OrdenMantenimiento.PRIORIDAD_ALTA],
        )
        .select_related("activo_ref")
        .order_by("fecha_programada", "id")[:20]
    )
    bitacora_mes_qs = BitacoraMantenimiento.objects.filter(fecha__date__gte=today - timedelta(days=30))

    costo_mes = (
        OrdenMantenimiento.objects.filter(
            fecha_cierre__year=today.year,
            fecha_cierre__month=today.month,
            estatus=OrdenMantenimiento.ESTATUS_CERRADA,
        ).aggregate(
            rep=Sum("costo_repuestos"),
            mo=Sum("costo_mano_obra"),
            otros=Sum("costo_otros"),
        )
    )
    costo_mes_total = _safe_decimal(costo_mes.get("rep")) + _safe_decimal(costo_mes.get("mo")) + _safe_decimal(
        costo_mes.get("otros")
    )

    criticidad_rows = list(
        activos_qs.values("criticidad")
        .annotate(total=Count("id"))
        .order_by()
    )
    criticidad = {
        "ALTA": 0,
        "MEDIA": 0,
        "BAJA": 0,
    }
    for row in criticidad_rows:
        criticidad[row["criticidad"]] = int(row["total"] or 0)

    active_plan_asset_ids = set(
        planes_activos_qs.values_list("activo_ref_id", flat=True)
    )
    overdue_plan_asset_ids = set(
        planes_activos_qs.filter(proxima_ejecucion__isnull=False, proxima_ejecucion__lt=today).values_list("activo_ref_id", flat=True)
    )
    critical_open_asset_ids = set(
        ordenes_abiertas_qs.filter(
            tipo=OrdenMantenimiento.TIPO_CORRECTIVO,
            prioridad__in=[OrdenMantenimiento.PRIORIDAD_CRITICA, OrdenMantenimiento.PRIORIDAD_ALTA],
        ).values_list("activo_ref_id", flat=True)
    )
    activos_blocker_rows = []
    for activo in activos_qs.select_related("proveedor_mantenimiento").order_by("nombre", "id")[:300]:
        profile = _activo_enterprise_profile(
            activo,
            overdue_plan_asset_ids=overdue_plan_asset_ids,
            active_plan_asset_ids=active_plan_asset_ids,
            critical_open_asset_ids=critical_open_asset_ids,
        )
        if profile["gaps"]:
            activos_blocker_rows.append(
                {
                    "activo": activo,
                    "status_label": profile["status_label"],
                    "status_tone": profile["status_tone"],
                    "gap_labels": ", ".join(g["label"] for g in profile["gaps"][:3]),
                    "next_action": profile["next_action"],
                }
            )

    enterprise_cards = [
        {
            "label": "Sin categoría",
            "count": activos_qs.filter(Q(categoria__isnull=True) | Q(categoria="")).count(),
            "tone": "warning",
            "url": "activos:activos?master_gap=SIN_CATEGORIA",
        },
        {
            "label": "Sin proveedor",
            "count": activos_qs.filter(proveedor_mantenimiento__isnull=True).count(),
            "tone": "warning",
            "url": "activos:activos?master_gap=SIN_PROVEEDOR",
        },
        {
            "label": "Sin plan",
            "count": max(activos_qs.count() - len(active_plan_asset_ids), 0),
            "tone": "warning",
            "url": "activos:activos?master_gap=SIN_PLAN",
        },
        {
            "label": "Plan vencido",
            "count": len(overdue_plan_asset_ids),
            "tone": "danger",
            "url": "activos:activos?master_gap=PLAN_VENCIDO",
        },
        {
            "label": "Orden crítica abierta",
            "count": len(critical_open_asset_ids),
            "tone": "danger",
            "url": "activos:activos?master_gap=ORDEN_CRITICA",
        },
        {
            "label": "Fuera de servicio crítico",
            "count": activos_qs.filter(
                estado=Activo.ESTADO_FUERA_SERVICIO,
                criticidad=Activo.CRITICIDAD_ALTA,
            ).count(),
            "tone": "danger",
            "url": "activos:activos?master_gap=FUERA_SERVICIO_CRITICO",
        },
    ]
    blockers_total = sum(card["count"] for card in enterprise_cards)
    enterprise_health_label = "Listo ERP" if blockers_total == 0 else "Con bloqueos"
    enterprise_health_tone = "success" if blockers_total == 0 else "danger"
    activos_listos_erp = max(activos_qs.count() - blockers_total, 0)

    enterprise_chain = _activos_enterprise_chain(
        activos_total=activos_qs.count(),
        blockers_total=blockers_total,
        planes_total=planes_activos_qs.count(),
        overdue_plans=len(overdue_plan_asset_ids),
        ordenes_abiertas=ordenes_abiertas_qs.count(),
        critical_open=len(critical_open_asset_ids),
        bitacora_30d=bitacora_mes_qs.count(),
    )
    enterprise_maturity_summary = _activos_maturity_summary(chain=enterprise_chain)
    enterprise_handoff_map = _activos_handoff_map(
        blockers_total=blockers_total,
        overdue_plans=len(overdue_plan_asset_ids),
        ordenes_abiertas=ordenes_abiertas_qs.count(),
        bitacora_30d=bitacora_mes_qs.count(),
    )
    document_stage_rows = _annotate_activos_document_stage_rows([
        {
            "label": "Activos listos ERP",
            "open": blockers_total,
            "closed": activos_listos_erp,
            "completion": round((activos_listos_erp / max(activos_qs.count(), 1)) * 100) if activos_qs.count() else 0,
            "detail": "Activos sin brechas vs activos con bloqueo abierto.",
            "next_step": "Cerrar maestro del activo",
            "url": reverse("activos:activos"),
        },
        {
            "label": "Planes preventivos",
            "open": len(overdue_plan_asset_ids),
            "closed": max(planes_activos_qs.count() - len(overdue_plan_asset_ids), 0),
            "completion": round(
                (max(planes_activos_qs.count() - len(overdue_plan_asset_ids), 0) / max(planes_activos_qs.count(), 1)) * 100
            )
            if planes_activos_qs.count()
            else 0,
            "detail": "Planes activos controlados contra planes vencidos.",
            "next_step": "Regularizar planes vencidos",
            "url": f"{reverse('activos:planes')}?scope=all",
        },
        {
            "label": "Órdenes de mantenimiento",
            "open": ordenes_abiertas_qs.count(),
            "closed": OrdenMantenimiento.objects.filter(estatus=OrdenMantenimiento.ESTATUS_CERRADA).count(),
            "completion": round(
                (
                    OrdenMantenimiento.objects.filter(estatus=OrdenMantenimiento.ESTATUS_CERRADA).count()
                    / max(OrdenMantenimiento.objects.count(), 1)
                )
                * 100
            )
            if OrdenMantenimiento.objects.count()
            else 0,
            "detail": "Órdenes abiertas vs cerradas con costo documentado.",
            "next_step": "Cerrar órdenes abiertas",
            "url": reverse("activos:ordenes"),
        },
        {
            "label": "Bitácora 30 días",
            "open": 0,
            "closed": bitacora_mes_qs.count(),
            "completion": 100 if bitacora_mes_qs.count() else 0,
            "detail": "Eventos registrados en el periodo reciente.",
            "next_step": "Mantener evidencia diaria",
            "url": reverse("activos:reportes"),
        },
    ])
    release_gate_rows = [
        {
            "step": "01",
            "title": "Activo maestro listo para operar",
            "detail": "Activos sin brechas críticas de catálogo, clasificación o criticidad.",
            "completed": activos_listos_erp,
            "open_count": blockers_total,
            "total": activos_qs.count(),
            "tone": "success" if blockers_total == 0 else "warning",
            "url": reverse("activos:activos"),
            "cta": "Abrir activos",
        },
        {
            "step": "02",
            "title": "Mantenimiento preventivo cubierto",
            "detail": "Planes preventivos activos sin vencimientos abiertos en la cadena documental.",
            "completed": max(planes_activos_qs.count() - len(overdue_plan_asset_ids), 0),
            "open_count": len(overdue_plan_asset_ids),
            "total": max(planes_activos_qs.count(), 1),
            "tone": "success" if len(overdue_plan_asset_ids) == 0 else "warning",
            "url": f"{reverse('activos:planes')}?scope=all",
            "cta": "Abrir planes",
        },
        {
            "step": "03",
            "title": "Órdenes y bitácora cerradas",
            "detail": "Órdenes con costo documentado y bitácora reciente para sostener auditoría operativa.",
            "completed": OrdenMantenimiento.objects.filter(estatus=OrdenMantenimiento.ESTATUS_CERRADA).count() + bitacora_mes_qs.count(),
            "open_count": ordenes_abiertas_qs.count(),
            "total": max(
                OrdenMantenimiento.objects.count() + bitacora_mes_qs.count(),
                1,
            ),
            "tone": "success" if ordenes_abiertas_qs.count() == 0 else "warning",
            "url": reverse("activos:ordenes"),
            "cta": "Abrir órdenes",
        },
    ]
    release_gate_completion = (
        round(
            (
                sum(row["completed"] for row in release_gate_rows)
                / sum(row["total"] for row in release_gate_rows)
            )
            * 100
        )
        if release_gate_rows and sum(row["total"] for row in release_gate_rows)
        else 0
    )
    workflow_stage_rows = [
        {
            "step": row["step"],
            "title": row["title"],
            "completion": round((row["completed"] / row["total"]) * 100) if row["total"] else 0,
            "detail": row["detail"],
            "owner": (
                "Activos / Mantenimiento"
                if row["step"] == 1
                else "Planeación / Operación"
                if row["step"] == 2
                else "Mantenimiento / Ejecución"
                if row["step"] == 3
                else "DG / Auditoría"
            ),
            "next_step": row["cta"],
            "url": row["url"],
            "tone": row["tone"],
        }
        for row in release_gate_rows
    ]

    context = {
        "module_tabs": _module_tabs("dashboard"),
        "activos_total": activos_qs.count(),
        "activos_operativos": activos_qs.filter(estado=Activo.ESTADO_OPERATIVO).count(),
        "activos_mantenimiento": activos_qs.filter(estado=Activo.ESTADO_MANTENIMIENTO).count(),
        "activos_fuera_servicio": activos_qs.filter(estado=Activo.ESTADO_FUERA_SERVICIO).count(),
        "ordenes_abiertas": ordenes_abiertas_qs.count(),
        "ordenes_en_proceso": ordenes_abiertas_qs.filter(estatus=OrdenMantenimiento.ESTATUS_EN_PROCESO).count(),
        "planes_vencidos": planes_activos_qs.filter(proxima_ejecucion__lt=today).count(),
        "planes_proxima_semana": planes_activos_qs.filter(
            proxima_ejecucion__gte=today,
            proxima_ejecucion__lte=week_limit,
        ).count(),
        "costo_mes_total": costo_mes_total,
        "criticidad": criticidad,
        "proximos": proximos,
        "ordenes_recientes": ordenes_recientes,
        "planes_vencidos_rows": planes_vencidos_rows,
        "ordenes_criticas_rows": ordenes_criticas_rows,
        "enterprise_cards": enterprise_cards,
        "enterprise_health_label": enterprise_health_label,
        "enterprise_health_tone": enterprise_health_tone,
        "enterprise_blockers_total": blockers_total,
        "enterprise_maturity_summary": enterprise_maturity_summary,
        "enterprise_handoff_map": enterprise_handoff_map,
        "operational_health_cards": _activos_operational_health_cards(
            cards=enterprise_cards,
            module_key="dashboard",
        ),
        "enterprise_chain": enterprise_chain,
        "document_stage_rows": document_stage_rows,
        "erp_governance_rows": _activos_erp_governance_rows(document_stage_rows),
        "activos_blocker_rows": activos_blocker_rows[:12],
        "release_gate_rows": release_gate_rows,
        "release_gate_completion": release_gate_completion,
        "workflow_stage_rows": workflow_stage_rows,
        "erp_command_center": _activos_command_center(
            owner="Activos / Mantenimiento",
            blockers_total=blockers_total,
            maturity_summary=enterprise_maturity_summary,
            default_url=reverse("activos:activos"),
            default_cta="Abrir activos",
        ),
    }
    context["critical_path_rows"] = _activos_critical_path_rows(context["enterprise_chain"])
    context["executive_radar_rows"] = _activos_executive_radar_rows(
        context["document_stage_rows"],
        context["enterprise_chain"],
    )
    return render(request, "activos/dashboard.html", context)


@login_required
def activos_catalog(request):
    if not can_view_inventario(request.user):
        raise PermissionDenied("No tienes permisos para ver Activos.")

    if request.method == "POST":
        if not can_manage_inventario(request.user):
            raise PermissionDenied("No tienes permisos para gestionar Activos.")
        action = (request.POST.get("action") or "create_activo").strip().lower()
        if action == "create_activo":
            nombre = (request.POST.get("nombre") or "").strip()
            if not nombre:
                messages.error(request, "Nombre del activo es obligatorio.")
                return redirect("activos:activos")
            estado = (request.POST.get("estado") or Activo.ESTADO_OPERATIVO).strip().upper()
            criticidad = (request.POST.get("criticidad") or Activo.CRITICIDAD_MEDIA).strip().upper()
            proveedor_id = _safe_int(request.POST.get("proveedor_mantenimiento_id"))
            activo = Activo.objects.create(
                nombre=nombre,
                categoria=(request.POST.get("categoria") or "").strip(),
                ubicacion=(request.POST.get("ubicacion") or "").strip(),
                estado=estado if estado in {x[0] for x in Activo.ESTADO_CHOICES} else Activo.ESTADO_OPERATIVO,
                criticidad=(
                    criticidad if criticidad in {x[0] for x in Activo.CRITICIDAD_CHOICES} else Activo.CRITICIDAD_MEDIA
                ),
                proveedor_mantenimiento_id=proveedor_id if proveedor_id > 0 else None,
                fecha_alta=_parse_date(request.POST.get("fecha_alta")) or timezone.localdate(),
                valor_reposicion=_safe_decimal(request.POST.get("valor_reposicion")),
                vida_util_meses=max(1, _safe_int(request.POST.get("vida_util_meses"), default=60)),
                horas_uso_promedio_mes=_safe_decimal(request.POST.get("horas_uso_promedio_mes")),
                notas=(request.POST.get("notas") or "").strip(),
                activo=(request.POST.get("activo") or "").strip().lower() in {"1", "on", "true", "yes"},
            )
            log_event(
                request.user,
                "CREATE",
                "activos.Activo",
                activo.id,
                {"codigo": activo.codigo, "nombre": activo.nombre, "estado": activo.estado},
            )
            messages.success(request, f"Activo {activo.codigo} creado.")
            return redirect("activos:activos")

        if action == "set_estado":
            activo_id = _safe_int(request.POST.get("activo_id"))
            activo_obj = get_object_or_404(Activo, pk=activo_id)
            estado = (request.POST.get("estado") or "").strip().upper()
            if estado in {x[0] for x in Activo.ESTADO_CHOICES} and estado != activo_obj.estado:
                from_estado = activo_obj.estado
                activo_obj.estado = estado
                activo_obj.save(update_fields=["estado", "actualizado_en"])
                log_event(
                    request.user,
                    "UPDATE",
                    "activos.Activo",
                    activo_obj.id,
                    {"from_estado": from_estado, "to_estado": estado},
                )
                messages.success(request, f"Estado actualizado: {activo_obj.codigo}.")
            return redirect("activos:activos")

        if action == "toggle_activo":
            activo_id = _safe_int(request.POST.get("activo_id"))
            activo_obj = get_object_or_404(Activo, pk=activo_id)
            activo_obj.activo = not activo_obj.activo
            activo_obj.save(update_fields=["activo", "actualizado_en"])
            log_event(
                request.user,
                "UPDATE",
                "activos.Activo",
                activo_obj.id,
                {"activo": activo_obj.activo},
            )
            messages.success(request, f"{activo_obj.codigo}: {'Activo' if activo_obj.activo else 'Inactivo'}.")
            return redirect("activos:activos")

        if action == "import_bitacora":
            archivo = request.FILES.get("archivo_bitacora")
            if not archivo:
                messages.error(request, "Selecciona un archivo XLSX o CSV para importar.")
                return redirect("activos:activos")
            is_dry_run = (request.POST.get("dry_run") or "").strip().lower() in {"1", "on", "true", "yes"}
            skip_servicios = (request.POST.get("skip_servicios") or "").strip().lower() in {"1", "on", "true", "yes"}
            try:
                stats = import_bitacora(
                    archivo,
                    sheet_name=(request.POST.get("sheet_name") or "").strip(),
                    dry_run=is_dry_run,
                    skip_servicios=skip_servicios,
                )
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect("activos:activos")
            except Exception:
                messages.error(
                    request,
                    "No se pudo procesar el archivo. Verifica formato de hoja/columnas (nombre, marca, modelo, serie, fechas y costos).",
                )
                return redirect("activos:activos")

            mode_label = "simulación (sin guardar)" if is_dry_run else "importación aplicada"
            log_event(
                request.user,
                "IMPORT",
                "activos.BitacoraImport",
                timezone.localtime().strftime("%Y%m%d%H%M%S"),
                {
                    "filename": getattr(archivo, "name", "bitacora"),
                    "dry_run": is_dry_run,
                    "skip_servicios": skip_servicios,
                    "sheet_name": stats.get("sheet_name", ""),
                    "source_format": stats.get("source_format", ""),
                    "filas_leidas": stats.get("filas_leidas", 0),
                    "filas_validas": stats.get("filas_validas", 0),
                    "activos_creados": stats.get("activos_creados", 0),
                    "activos_actualizados": stats.get("activos_actualizados", 0),
                    "servicios_creados": stats.get("servicios_creados", 0),
                    "servicios_omitidos": stats.get("servicios_omitidos", 0),
                },
            )
            messages.success(
                request,
                (
                    f"Bitácora procesada ({mode_label}): filas válidas {stats['filas_validas']}, "
                    f"activos creados {stats['activos_creados']}, actualizados {stats['activos_actualizados']}, "
                    f"servicios creados {stats['servicios_creados']}, omitidos {stats['servicios_omitidos']}."
                ),
            )
            return redirect("activos:activos")

        messages.error(request, "Acción no reconocida.")
        return redirect("activos:activos")

    q = (request.GET.get("q") or "").strip()
    estado = (request.GET.get("estado") or "").strip().upper()
    criticidad = (request.GET.get("criticidad") or "").strip().upper()
    solo_activos = (request.GET.get("solo_activos") or "1").strip()
    master_gap = (request.GET.get("master_gap") or "").strip().upper()
    import_q = (request.GET.get("import_q") or "").strip()
    import_mode = (request.GET.get("import_mode") or "ALL").strip().upper()
    import_format = (request.GET.get("import_format") or "ALL").strip().upper()

    today = timezone.localdate()
    all_activos_qs = Activo.objects.select_related("proveedor_mantenimiento").order_by("nombre", "id")
    all_planes_qs = PlanMantenimiento.objects.filter(estatus=PlanMantenimiento.ESTATUS_ACTIVO, activo=True)
    all_ordenes_qs = OrdenMantenimiento.objects.filter(
        estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO]
    )
    bitacora_30d_qs = BitacoraMantenimiento.objects.filter(fecha__date__gte=today - timedelta(days=30))
    active_plan_asset_ids = set(all_planes_qs.values_list("activo_ref_id", flat=True))
    overdue_plan_asset_ids = set(
        all_planes_qs.filter(
            proxima_ejecucion__isnull=False,
            proxima_ejecucion__lt=today,
        ).values_list("activo_ref_id", flat=True)
    )
    critical_open_asset_ids = set(
        all_ordenes_qs.filter(
            estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO],
            tipo=OrdenMantenimiento.TIPO_CORRECTIVO,
            prioridad__in=[OrdenMantenimiento.PRIORIDAD_CRITICA, OrdenMantenimiento.PRIORIDAD_ALTA],
        ).values_list("activo_ref_id", flat=True)
    )

    qs = all_activos_qs
    if q:
        qs = qs.filter(Q(codigo__icontains=q) | Q(nombre__icontains=q) | Q(categoria__icontains=q) | Q(ubicacion__icontains=q))
    if estado in {x[0] for x in Activo.ESTADO_CHOICES}:
        qs = qs.filter(estado=estado)
    if criticidad in {x[0] for x in Activo.CRITICIDAD_CHOICES}:
        qs = qs.filter(criticidad=criticidad)
    if solo_activos == "1":
        qs = qs.filter(activo=True)
    if master_gap == "SIN_CATEGORIA":
        qs = qs.filter(Q(categoria__isnull=True) | Q(categoria=""))
    elif master_gap == "SIN_PROVEEDOR":
        qs = qs.filter(proveedor_mantenimiento__isnull=True)
    elif master_gap == "SIN_PLAN":
        qs = qs.exclude(id__in=active_plan_asset_ids)
    elif master_gap == "PLAN_VENCIDO":
        qs = qs.filter(id__in=overdue_plan_asset_ids)
    elif master_gap == "ORDEN_CRITICA":
        qs = qs.filter(id__in=critical_open_asset_ids)
    elif master_gap == "FUERA_SERVICIO_CRITICO":
        qs = qs.filter(estado=Activo.ESTADO_FUERA_SERVICIO, criticidad=Activo.CRITICIDAD_ALTA)

    export_format = (request.GET.get("export") or "").strip().lower()
    import_runs_qs = (
        AuditLog.objects.select_related("user")
        .filter(action="IMPORT", model="activos.BitacoraImport")
        .order_by("-timestamp")
    )
    if import_mode in {"APLICADO", "SIMULACION"}:
        target_dry_run = import_mode == "SIMULACION"
        import_runs_qs = import_runs_qs.filter(payload__dry_run=target_dry_run)
    if import_format in {"CSV", "XLSX"}:
        import_runs_qs = import_runs_qs.filter(payload__source_format=import_format)
    if import_q:
        import_runs_qs = import_runs_qs.filter(
            Q(payload__filename__icontains=import_q)
            | Q(user__username__icontains=import_q)
        )
    if export_format == "template_bitacora_csv":
        return _export_bitacora_template_csv()
    if export_format == "template_bitacora_xlsx":
        return _export_bitacora_template_xlsx()
    if export_format == "import_runs_csv":
        return _export_bitacora_runs_csv(import_runs_qs[:300])
    if export_format == "import_runs_xlsx":
        return _export_bitacora_runs_xlsx(import_runs_qs[:300])

    if export_format in {"depuracion_csv", "depuracion_xlsx"}:
        all_name_counts = {
            (row["nombre_lower"] or ""): int(row["total"] or 0)
            for row in Activo.objects.filter(activo=True)
            .annotate(nombre_lower=Lower("nombre"))
            .values("nombre_lower")
            .annotate(total=Count("id"))
        }
        dep_rows = _build_activos_depuracion_rows(list(qs), all_name_counts=all_name_counts)
        if export_format == "depuracion_csv":
            return _export_activos_depuracion_csv(dep_rows)
        return _export_activos_depuracion_xlsx(dep_rows)

    activos_rows = []
    for activo in list(qs[:300]):
        profile = _activo_enterprise_profile(
            activo,
            overdue_plan_asset_ids=overdue_plan_asset_ids,
            active_plan_asset_ids=active_plan_asset_ids,
            critical_open_asset_ids=critical_open_asset_ids,
        )
        activos_rows.append(
            {
                "activo": activo,
                "enterprise": profile,
            }
        )

    enterprise_cards = [
        {
            "key": "SIN_CATEGORIA",
            "label": "Sin categoría",
            "count": all_activos_qs.filter(Q(categoria__isnull=True) | Q(categoria="")).count(),
            "tone": "warning",
        },
        {
            "key": "SIN_PROVEEDOR",
            "label": "Sin proveedor",
            "count": all_activos_qs.filter(proveedor_mantenimiento__isnull=True).count(),
            "tone": "warning",
        },
        {
            "key": "SIN_PLAN",
            "label": "Sin plan",
            "count": max(all_activos_qs.count() - len(active_plan_asset_ids), 0),
            "tone": "warning",
        },
        {
            "key": "PLAN_VENCIDO",
            "label": "Plan vencido",
            "count": len(overdue_plan_asset_ids),
            "tone": "danger",
        },
        {
            "key": "ORDEN_CRITICA",
            "label": "Orden crítica abierta",
            "count": len(critical_open_asset_ids),
            "tone": "danger",
        },
        {
            "key": "FUERA_SERVICIO_CRITICO",
            "label": "Fuera de servicio crítico",
            "count": all_activos_qs.filter(
                estado=Activo.ESTADO_FUERA_SERVICIO,
                criticidad=Activo.CRITICIDAD_ALTA,
            ).count(),
            "tone": "danger",
        },
    ]

    blockers_total = sum(card["count"] for card in enterprise_cards)
    enterprise_chain = _activos_enterprise_chain(
        activos_total=all_activos_qs.count(),
        blockers_total=blockers_total,
        planes_total=all_planes_qs.count(),
        overdue_plans=len(overdue_plan_asset_ids),
        ordenes_abiertas=all_ordenes_qs.count(),
        critical_open=len(critical_open_asset_ids),
        bitacora_30d=bitacora_30d_qs.count(),
    )
    enterprise_maturity_summary = _activos_maturity_summary(chain=enterprise_chain)
    enterprise_handoff_map = _activos_handoff_map(
        blockers_total=blockers_total,
        overdue_plans=len(overdue_plan_asset_ids),
        ordenes_abiertas=all_ordenes_qs.count(),
        bitacora_30d=bitacora_30d_qs.count(),
    )
    document_stage_rows = _annotate_activos_document_stage_rows([
        {
            "label": "Activos catalogados",
            "open": blockers_total,
            "closed": max(all_activos_qs.count() - blockers_total, 0),
            "detail": "Activos sin clasificación completa, proveedor o plan vs activos listos.",
            "url": reverse("activos:activos"),
        },
        {
            "label": "Planes preventivos",
            "open": len(overdue_plan_asset_ids),
            "closed": max(all_planes_qs.count() - len(overdue_plan_asset_ids), 0),
            "detail": "Planes activos controlados contra planes vencidos.",
            "url": reverse("activos:planes"),
        },
        {
            "label": "Órdenes documentales",
            "open": all_ordenes_qs.count(),
            "closed": bitacora_30d_qs.count(),
            "detail": "Órdenes abiertas frente a eventos y cierres registrados.",
            "url": reverse("activos:ordenes"),
        },
        {
            "label": "Bitácora y trazabilidad",
            "open": 0,
            "closed": bitacora_30d_qs.count(),
            "detail": "Eventos registrados en los últimos 30 días.",
            "url": reverse("activos:reportes"),
        },
    ])

    context = {
        "module_tabs": _module_tabs("activos"),
        "activos_rows": activos_rows,
        "proveedores": list(Proveedor.objects.filter(activo=True).order_by("nombre")[:800]),
        "estado_choices": Activo.ESTADO_CHOICES,
        "criticidad_choices": Activo.CRITICIDAD_CHOICES,
        "filters": {"q": q, "estado": estado, "criticidad": criticidad, "solo_activos": solo_activos, "master_gap": master_gap},
        "import_filters": {"import_q": import_q, "import_mode": import_mode, "import_format": import_format},
        "can_manage_activos": can_manage_inventario(request.user),
        "import_runs": list(import_runs_qs[:10]),
        "enterprise_cards": enterprise_cards,
        "enterprise_chain": enterprise_chain,
        "enterprise_maturity_summary": enterprise_maturity_summary,
        "enterprise_handoff_map": enterprise_handoff_map,
        "document_stage_rows": document_stage_rows,
        "erp_governance_rows": _activos_erp_governance_rows(document_stage_rows),
        "enterprise_focus_cards": _activos_focus_cards(
            selected_focus=master_gap,
            enterprise_cards=enterprise_cards,
            module_key="activos",
        ),
        "focus_summary": (
            _active_focus_summary(
                focus_label=next(card["label"] for card in enterprise_cards if card["key"] == master_gap),
                focus_detail=f"Vista filtrada por {next(card['label'].lower() for card in enterprise_cards if card['key'] == master_gap)} para cierre del maestro.",
                clear_url=reverse("activos:activos"),
            )
            if master_gap and any(card["key"] == master_gap for card in enterprise_cards)
            else None
        ),
        "erp_command_center": _activos_command_center(
            owner="Activos / Mantenimiento",
            blockers_total=blockers_total,
            maturity_summary=enterprise_maturity_summary,
            default_url=reverse("activos:activos"),
            default_cta="Abrir catálogo",
        ),
    }
    context["release_gate_rows"] = _activos_release_gate_rows(context["document_stage_rows"])
    context["release_gate_completion"] = _activos_release_gate_completion(context["release_gate_rows"])
    context["critical_path_rows"] = _activos_critical_path_rows(context["enterprise_chain"])
    context["executive_radar_rows"] = _activos_executive_radar_rows(
        context["document_stage_rows"],
        context["enterprise_chain"],
    )
    return render(request, "activos/activos.html", context)


@login_required
def planes(request):
    if not can_view_inventario(request.user):
        raise PermissionDenied("No tienes permisos para ver Activos.")

    if request.method == "POST":
        if not can_manage_inventario(request.user):
            raise PermissionDenied("No tienes permisos para gestionar planes.")
        action = (request.POST.get("action") or "create_plan").strip().lower()
        if action == "create_plan":
            activo_id = _safe_int(request.POST.get("activo_id"))
            nombre = (request.POST.get("nombre") or "").strip()
            if not activo_id or not nombre:
                messages.error(request, "Activo y nombre del plan son obligatorios.")
                return redirect("activos:planes")
            activo_obj = get_object_or_404(Activo, pk=activo_id)
            tipo = (request.POST.get("tipo") or PlanMantenimiento.TIPO_PREVENTIVO).strip().upper()
            estatus = (request.POST.get("estatus") or PlanMantenimiento.ESTATUS_ACTIVO).strip().upper()
            plan = PlanMantenimiento.objects.create(
                activo_ref=activo_obj,
                nombre=nombre,
                tipo=tipo if tipo in {x[0] for x in PlanMantenimiento.TIPO_CHOICES} else PlanMantenimiento.TIPO_PREVENTIVO,
                frecuencia_dias=max(1, _safe_int(request.POST.get("frecuencia_dias"), default=30)),
                tolerancia_dias=max(0, _safe_int(request.POST.get("tolerancia_dias"), default=0)),
                ultima_ejecucion=_parse_date(request.POST.get("ultima_ejecucion")),
                proxima_ejecucion=_parse_date(request.POST.get("proxima_ejecucion")),
                responsable=(request.POST.get("responsable") or "").strip(),
                instrucciones=(request.POST.get("instrucciones") or "").strip(),
                estatus=(
                    estatus if estatus in {x[0] for x in PlanMantenimiento.ESTATUS_CHOICES} else PlanMantenimiento.ESTATUS_ACTIVO
                ),
                activo=(request.POST.get("activo") or "").strip().lower() in {"1", "on", "true", "yes"},
            )
            log_event(
                request.user,
                "CREATE",
                "activos.PlanMantenimiento",
                plan.id,
                {"activo_id": plan.activo_ref_id, "nombre": plan.nombre, "proxima_ejecucion": str(plan.proxima_ejecucion or "")},
            )
            messages.success(request, f"Plan creado para {activo_obj.nombre}.")
            return redirect("activos:planes")

        if action == "toggle_plan":
            plan_id = _safe_int(request.POST.get("plan_id"))
            plan = get_object_or_404(PlanMantenimiento, pk=plan_id)
            plan.estatus = (
                PlanMantenimiento.ESTATUS_PAUSADO
                if plan.estatus == PlanMantenimiento.ESTATUS_ACTIVO
                else PlanMantenimiento.ESTATUS_ACTIVO
            )
            plan.save(update_fields=["estatus", "actualizado_en"])
            log_event(request.user, "UPDATE", "activos.PlanMantenimiento", plan.id, {"estatus": plan.estatus})
            messages.success(request, f"Plan {plan.nombre} actualizado.")
            return redirect("activos:planes")

        if action == "registrar_ejecucion":
            plan_id = _safe_int(request.POST.get("plan_id"))
            plan = get_object_or_404(PlanMantenimiento, pk=plan_id)
            fecha = _parse_date(request.POST.get("fecha")) or timezone.localdate()
            plan.ultima_ejecucion = fecha
            plan.recompute_next_date()
            plan.save(update_fields=["ultima_ejecucion", "proxima_ejecucion", "actualizado_en"])
            log_event(
                request.user,
                "UPDATE",
                "activos.PlanMantenimiento",
                plan.id,
                {"ultima_ejecucion": str(fecha), "proxima_ejecucion": str(plan.proxima_ejecucion or "")},
            )
            messages.success(request, f"Ejecución registrada para {plan.nombre}.")
            return redirect("activos:planes")

        if action == "generar_ordenes_programadas":
            scope = (request.POST.get("scope") or "overdue").strip().lower()
            dry_run = (request.POST.get("dry_run") or "").strip().lower() in {"1", "on", "true", "yes"}
            today = timezone.localdate()
            if scope == "week":
                plan_qs = PlanMantenimiento.objects.filter(
                    estatus=PlanMantenimiento.ESTATUS_ACTIVO,
                    activo=True,
                    proxima_ejecucion__isnull=False,
                    proxima_ejecucion__gte=today,
                    proxima_ejecucion__lte=today + timedelta(days=7),
                )
            else:
                plan_qs = PlanMantenimiento.objects.filter(
                    estatus=PlanMantenimiento.ESTATUS_ACTIVO,
                    activo=True,
                    proxima_ejecucion__isnull=False,
                    proxima_ejecucion__lte=today,
                )

            created = 0
            skipped = 0
            plan_qs = plan_qs.select_related("activo_ref").order_by("proxima_ejecucion", "id")
            for plan in plan_qs:
                if not plan.activo_ref or not plan.activo_ref.activo:
                    skipped += 1
                    continue

                exists = OrdenMantenimiento.objects.filter(
                    plan_ref=plan,
                    fecha_programada=plan.proxima_ejecucion,
                ).exclude(estatus=OrdenMantenimiento.ESTATUS_CANCELADA).exists()
                if exists:
                    skipped += 1
                    continue

                if dry_run:
                    created += 1
                    continue

                orden = OrdenMantenimiento.objects.create(
                    activo_ref=plan.activo_ref,
                    plan_ref=plan,
                    tipo=OrdenMantenimiento.TIPO_PREVENTIVO,
                    prioridad=_prioridad_por_criticidad(plan.activo_ref.criticidad),
                    estatus=OrdenMantenimiento.ESTATUS_PENDIENTE,
                    fecha_programada=plan.proxima_ejecucion or today,
                    responsable=plan.responsable or "",
                    descripcion=f"Orden preventiva automática desde plan: {plan.nombre}",
                    creado_por=request.user if request.user.is_authenticated else None,
                )
                BitacoraMantenimiento.objects.create(
                    orden=orden,
                    accion="AUTO_PLAN",
                    comentario="Generada automáticamente desde plan activo",
                    usuario=request.user if request.user.is_authenticated else None,
                    costo_adicional=Decimal("0"),
                )
                created += 1
                log_event(
                    request.user,
                    "CREATE",
                    "activos.OrdenMantenimiento",
                    orden.id,
                    {
                        "origen": "plan_auto",
                        "plan_id": plan.id,
                        "fecha_programada": str(plan.proxima_ejecucion or ""),
                        "folio": orden.folio,
                    },
                )

            run_mode = "simulación" if dry_run else "aplicado"
            messages.success(
                request,
                f"Generación de órdenes ({run_mode}): creadas {created}, omitidas {skipped}.",
            )
            return redirect("activos:planes")

        messages.error(request, "Acción no reconocida.")
        return redirect("activos:planes")

    q = (request.GET.get("q") or "").strip()
    estatus = (request.GET.get("estatus") or "").strip().upper()
    scope = (request.GET.get("scope") or "all").strip().lower()
    enterprise_gap = (request.GET.get("enterprise_gap") or "").strip().upper()
    today = timezone.localdate()

    qs = PlanMantenimiento.objects.select_related("activo_ref").order_by("proxima_ejecucion", "id")
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(activo_ref__nombre__icontains=q) | Q(activo_ref__codigo__icontains=q))
    if estatus in {x[0] for x in PlanMantenimiento.ESTATUS_CHOICES}:
        qs = qs.filter(estatus=estatus)
    if scope == "overdue":
        qs = qs.filter(proxima_ejecucion__isnull=False, proxima_ejecucion__lt=today, estatus=PlanMantenimiento.ESTATUS_ACTIVO)
    elif scope == "week":
        qs = qs.filter(
            proxima_ejecucion__isnull=False,
            proxima_ejecucion__gte=today,
            proxima_ejecucion__lte=today + timedelta(days=7),
            estatus=PlanMantenimiento.ESTATUS_ACTIVO,
        )
    if enterprise_gap == "SIN_RESPONSABLE":
        qs = qs.filter(Q(responsable__isnull=True) | Q(responsable=""))
    elif enterprise_gap == "SIN_PROXIMA":
        qs = qs.filter(proxima_ejecucion__isnull=True)
    elif enterprise_gap == "VENCIDO":
        qs = qs.filter(proxima_ejecucion__isnull=False, proxima_ejecucion__lt=today, estatus=PlanMantenimiento.ESTATUS_ACTIVO)
    elif enterprise_gap == "ACTIVO_FUERA_SERVICIO":
        qs = qs.filter(activo_ref__estado=Activo.ESTADO_FUERA_SERVICIO)

    export_format = (request.GET.get("export") or "").strip().lower()
    if export_format in {"csv", "xlsx"}:
        export_rows = list(qs[:1500])
        if export_format == "csv":
            return _export_planes_csv(export_rows)
        return _export_planes_xlsx(export_rows)

    planes_rows = []
    for plan in list(qs[:300]):
        planes_rows.append({"plan": plan, "enterprise": _plan_enterprise_profile(plan, today=today)})
    all_planes_qs = PlanMantenimiento.objects.select_related("activo_ref")
    enterprise_cards = [
        {"key": "SIN_RESPONSABLE", "label": "Sin responsable", "count": all_planes_qs.filter(Q(responsable__isnull=True) | Q(responsable="")).count(), "tone": "warning"},
        {"key": "SIN_PROXIMA", "label": "Sin próxima", "count": all_planes_qs.filter(proxima_ejecucion__isnull=True).count(), "tone": "warning"},
        {"key": "VENCIDO", "label": "Vencidos", "count": all_planes_qs.filter(proxima_ejecucion__isnull=False, proxima_ejecucion__lt=today, estatus=PlanMantenimiento.ESTATUS_ACTIVO).count(), "tone": "danger"},
        {"key": "ACTIVO_FUERA_SERVICIO", "label": "Activo fuera servicio", "count": all_planes_qs.filter(activo_ref__estado=Activo.ESTADO_FUERA_SERVICIO).count(), "tone": "danger"},
    ]
    activos_total = Activo.objects.filter(activo=True).count()
    active_plan_asset_ids = set(
        all_planes_qs.filter(estatus=PlanMantenimiento.ESTATUS_ACTIVO, activo=True).values_list("activo_ref_id", flat=True)
    )
    overdue_plan_asset_ids = set(
        all_planes_qs.filter(
            estatus=PlanMantenimiento.ESTATUS_ACTIVO,
            activo=True,
            proxima_ejecucion__isnull=False,
            proxima_ejecucion__lt=today,
        ).values_list("activo_ref_id", flat=True)
    )
    critical_open_asset_ids = set(
        OrdenMantenimiento.objects.filter(
            estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO],
            tipo=OrdenMantenimiento.TIPO_CORRECTIVO,
            prioridad__in=[OrdenMantenimiento.PRIORIDAD_CRITICA, OrdenMantenimiento.PRIORIDAD_ALTA],
        ).values_list("activo_ref_id", flat=True)
    )
    blockers_total = (
        Activo.objects.filter(Q(categoria__isnull=True) | Q(categoria=""), activo=True).count()
        + Activo.objects.filter(proveedor_mantenimiento__isnull=True, activo=True).count()
        + max(activos_total - len(active_plan_asset_ids), 0)
        + len(overdue_plan_asset_ids)
        + len(critical_open_asset_ids)
        + Activo.objects.filter(
            estado=Activo.ESTADO_FUERA_SERVICIO,
            criticidad=Activo.CRITICIDAD_ALTA,
            activo=True,
        ).count()
    )
    enterprise_chain = _activos_enterprise_chain(
        activos_total=activos_total,
        blockers_total=blockers_total,
        planes_total=all_planes_qs.filter(estatus=PlanMantenimiento.ESTATUS_ACTIVO, activo=True).count(),
        overdue_plans=len(overdue_plan_asset_ids),
        ordenes_abiertas=OrdenMantenimiento.objects.filter(
            estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO]
        ).count(),
        critical_open=len(critical_open_asset_ids),
        bitacora_30d=BitacoraMantenimiento.objects.filter(fecha__date__gte=today - timedelta(days=30)).count(),
    )
    enterprise_maturity_summary = _activos_maturity_summary(chain=enterprise_chain)
    enterprise_handoff_map = _activos_handoff_map(
        blockers_total=blockers_total,
        overdue_plans=len(overdue_plan_asset_ids),
        ordenes_abiertas=OrdenMantenimiento.objects.filter(
            estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO]
        ).count(),
        bitacora_30d=BitacoraMantenimiento.objects.filter(fecha__date__gte=today - timedelta(days=30)).count(),
    )
    context = {
        "module_tabs": _module_tabs("planes"),
        "planes_rows": planes_rows,
        "activos": list(Activo.objects.filter(activo=True).order_by("nombre")[:800]),
        "tipo_choices": PlanMantenimiento.TIPO_CHOICES,
        "estatus_choices": PlanMantenimiento.ESTATUS_CHOICES,
        "filters": {"q": q, "estatus": estatus, "scope": scope, "enterprise_gap": enterprise_gap},
        "today": today,
        "can_manage_activos": can_manage_inventario(request.user),
        "enterprise_cards": enterprise_cards,
        "operational_health_cards": _activos_operational_health_cards(cards=enterprise_cards, module_key="planes"),
        "enterprise_chain": enterprise_chain,
        "enterprise_maturity_summary": enterprise_maturity_summary,
        "enterprise_handoff_map": enterprise_handoff_map,
        "document_stage_rows": _annotate_activos_document_stage_rows([
            {
                "label": "Planes activos",
                "open": enterprise_cards[0]["count"] + enterprise_cards[1]["count"],
                "closed": max(all_planes_qs.filter(estatus=PlanMantenimiento.ESTATUS_ACTIVO, activo=True).count() - (enterprise_cards[0]["count"] + enterprise_cards[1]["count"]), 0),
                "detail": "Planes sin responsable o sin próxima ejecución contra planes controlados.",
                "url": reverse("activos:planes"),
            },
            {
                "label": "Planes vencidos",
                "open": enterprise_cards[2]["count"],
                "closed": max(all_planes_qs.count() - enterprise_cards[2]["count"], 0),
                "detail": "Planes fuera de fecha frente a planes en calendario.",
                "url": reverse("activos:planes"),
            },
            {
                "label": "Activos condicionantes",
                "open": enterprise_cards[3]["count"],
                "closed": max(activos_total - enterprise_cards[3]["count"], 0),
                "detail": "Activos fuera de servicio que bloquean la disciplina preventiva.",
                "url": reverse("activos:activos"),
            },
            {
                "label": "Órdenes preventivas",
                "open": OrdenMantenimiento.objects.filter(
                    tipo=OrdenMantenimiento.TIPO_PREVENTIVO,
                    estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO],
                ).count(),
                "closed": BitacoraMantenimiento.objects.filter(
                    orden__tipo=OrdenMantenimiento.TIPO_PREVENTIVO,
                    fecha__date__gte=today - timedelta(days=30),
                ).count(),
                "detail": "Órdenes preventivas abiertas frente a cierres registrados.",
                "url": reverse("activos:ordenes"),
            },
        ]),
        "erp_command_center": _activos_command_center(
            owner="Planeación / Operación",
            blockers_total=sum(card["count"] for card in enterprise_cards),
            maturity_summary=enterprise_maturity_summary,
            default_url=reverse("activos:planes"),
            default_cta="Abrir planes",
        ),
    }
    context["erp_governance_rows"] = _activos_erp_governance_rows(context["document_stage_rows"])
    context["release_gate_rows"] = _activos_release_gate_rows(context["document_stage_rows"])
    context["release_gate_completion"] = _activos_release_gate_completion(context["release_gate_rows"])
    context["critical_path_rows"] = _activos_critical_path_rows(context["enterprise_chain"])
    context["executive_radar_rows"] = _activos_executive_radar_rows(
        context["document_stage_rows"],
        context["enterprise_chain"],
    )
    return render(request, "activos/planes.html", context)


@login_required
def ordenes(request):
    if not can_view_inventario(request.user):
        raise PermissionDenied("No tienes permisos para ver Activos.")

    if request.method == "POST":
        if not can_manage_inventario(request.user):
            raise PermissionDenied("No tienes permisos para gestionar órdenes de mantenimiento.")
        action = (request.POST.get("action") or "create_orden").strip().lower()
        if action == "create_orden":
            activo_id = (request.POST.get("activo_id") or "").strip()
            plan_id = (request.POST.get("plan_id") or "").strip()
            tipo = (request.POST.get("tipo") or OrdenMantenimiento.TIPO_PREVENTIVO).strip().upper()
            prioridad = (request.POST.get("prioridad") or OrdenMantenimiento.PRIORIDAD_MEDIA).strip().upper()
            descripcion = (request.POST.get("descripcion") or "").strip()
            responsable = (request.POST.get("responsable") or "").strip()
            fecha_programada_raw = (request.POST.get("fecha_programada") or "").strip()
            try:
                fecha_programada = (
                    timezone.datetime.fromisoformat(fecha_programada_raw).date()
                    if fecha_programada_raw
                    else timezone.localdate()
                )
            except ValueError:
                fecha_programada = timezone.localdate()
            if not activo_id.isdigit():
                messages.error(request, "Selecciona un activo válido.")
                return redirect("activos:ordenes")
            activo_obj = get_object_or_404(Activo, pk=int(activo_id))
            plan_obj = None
            if plan_id.isdigit():
                plan_obj = PlanMantenimiento.objects.filter(pk=int(plan_id), activo_ref=activo_obj).first()
            orden = OrdenMantenimiento.objects.create(
                activo_ref=activo_obj,
                plan_ref=plan_obj,
                tipo=tipo if tipo in {x[0] for x in OrdenMantenimiento.TIPO_CHOICES} else OrdenMantenimiento.TIPO_PREVENTIVO,
                prioridad=(
                    prioridad
                    if prioridad in {x[0] for x in OrdenMantenimiento.PRIORIDAD_CHOICES}
                    else OrdenMantenimiento.PRIORIDAD_MEDIA
                ),
                descripcion=descripcion,
                responsable=responsable,
                fecha_programada=fecha_programada,
                creado_por=request.user,
            )
            BitacoraMantenimiento.objects.create(
                orden=orden,
                accion="CREADA",
                comentario="Orden creada desde UI",
                usuario=request.user,
            )
            log_event(
                request.user,
                "CREATE",
                "activos.OrdenMantenimiento",
                orden.id,
                {
                    "folio": orden.folio,
                    "activo_id": orden.activo_ref_id,
                    "tipo": orden.tipo,
                    "prioridad": orden.prioridad,
                    "estatus": orden.estatus,
                },
            )
            messages.success(request, f"Orden {orden.folio} creada.")
            return redirect("activos:ordenes")

        if action == "update_costos":
            orden_id = _safe_int(request.POST.get("orden_id"))
            if not orden_id:
                messages.error(request, "Selecciona una orden válida.")
                return redirect("activos:ordenes")
            orden = get_object_or_404(OrdenMantenimiento, pk=orden_id)
            orden.costo_repuestos = _safe_decimal(request.POST.get("costo_repuestos"))
            orden.costo_mano_obra = _safe_decimal(request.POST.get("costo_mano_obra"))
            orden.costo_otros = _safe_decimal(request.POST.get("costo_otros"))
            close_now = (request.POST.get("cerrar_orden") or "").strip().lower() in {"1", "on", "true", "yes"}
            if close_now and orden.estatus != OrdenMantenimiento.ESTATUS_CERRADA:
                orden.estatus = OrdenMantenimiento.ESTATUS_CERRADA
                if not orden.fecha_inicio:
                    orden.fecha_inicio = timezone.localdate()
                orden.fecha_cierre = timezone.localdate()
            orden.save(
                update_fields=[
                    "costo_repuestos",
                    "costo_mano_obra",
                    "costo_otros",
                    "estatus",
                    "fecha_inicio",
                    "fecha_cierre",
                    "actualizado_en",
                ]
            )
            BitacoraMantenimiento.objects.create(
                orden=orden,
                accion="COSTOS",
                comentario=(
                    f"Costos actualizados: repuestos={orden.costo_repuestos}, "
                    f"mano_obra={orden.costo_mano_obra}, otros={orden.costo_otros}"
                ),
                usuario=request.user,
            )
            log_event(
                request.user,
                "UPDATE",
                "activos.OrdenMantenimiento",
                orden.id,
                {
                    "folio": orden.folio,
                    "costo_repuestos": str(orden.costo_repuestos),
                    "costo_mano_obra": str(orden.costo_mano_obra),
                    "costo_otros": str(orden.costo_otros),
                    "estatus": orden.estatus,
                },
            )
            messages.success(request, f"Orden {orden.folio} actualizada (costos).")
            return redirect("activos:ordenes")

        messages.error(request, "Acción no reconocida.")
        return redirect("activos:ordenes")

    estado = (request.GET.get("estatus") or "abiertas").strip().upper()
    enterprise_gap = (request.GET.get("enterprise_gap") or "").strip().upper()
    qs = OrdenMantenimiento.objects.select_related("activo_ref", "plan_ref", "creado_por").order_by("-fecha_programada", "-id")
    if estado == "ABIERTAS":
        qs = qs.filter(estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO])
    elif estado in {x[0] for x in OrdenMantenimiento.ESTATUS_CHOICES}:
        qs = qs.filter(estatus=estado)
    if enterprise_gap == "SIN_RESPONSABLE":
        qs = qs.filter(Q(responsable__isnull=True) | Q(responsable=""))
    elif enterprise_gap == "SIN_FECHA_INICIO":
        qs = qs.filter(estatus=OrdenMantenimiento.ESTATUS_EN_PROCESO, fecha_inicio__isnull=True)
    elif enterprise_gap == "SIN_COSTO_CIERRE":
        qs = qs.filter(
            estatus=OrdenMantenimiento.ESTATUS_CERRADA,
            costo_repuestos=Decimal("0"),
            costo_mano_obra=Decimal("0"),
            costo_otros=Decimal("0"),
        )
    elif enterprise_gap == "SIN_PLAN_ORIGEN":
        qs = qs.filter(tipo=OrdenMantenimiento.TIPO_PREVENTIVO, plan_ref__isnull=True)
    elif enterprise_gap == "ABIERTA_CRITICA":
        qs = qs.filter(
            prioridad__in=[OrdenMantenimiento.PRIORIDAD_CRITICA, OrdenMantenimiento.PRIORIDAD_ALTA],
            estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO],
        )

    export_format = (request.GET.get("export") or "").strip().lower()
    if export_format in {"csv", "xlsx"}:
        export_rows = list(qs[:2000])
        if export_format == "csv":
            return _export_ordenes_csv(export_rows)
        return _export_ordenes_xlsx(export_rows)

    ordenes_rows = []
    for orden in list(qs[:120]):
        ordenes_rows.append({"orden": orden, "enterprise": _orden_enterprise_profile(orden)})
    all_ordenes_qs = OrdenMantenimiento.objects.select_related("activo_ref", "plan_ref", "creado_por")
    enterprise_cards = [
        {"key": "SIN_RESPONSABLE", "label": "Sin responsable", "count": all_ordenes_qs.filter(Q(responsable__isnull=True) | Q(responsable="")).count(), "tone": "warning"},
        {"key": "SIN_FECHA_INICIO", "label": "En proceso sin inicio", "count": all_ordenes_qs.filter(estatus=OrdenMantenimiento.ESTATUS_EN_PROCESO, fecha_inicio__isnull=True).count(), "tone": "warning"},
        {"key": "SIN_COSTO_CIERRE", "label": "Cierre sin costo", "count": all_ordenes_qs.filter(estatus=OrdenMantenimiento.ESTATUS_CERRADA, costo_repuestos=Decimal("0"), costo_mano_obra=Decimal("0"), costo_otros=Decimal("0")).count(), "tone": "warning"},
        {"key": "SIN_PLAN_ORIGEN", "label": "Preventiva sin plan", "count": all_ordenes_qs.filter(tipo=OrdenMantenimiento.TIPO_PREVENTIVO, plan_ref__isnull=True).count(), "tone": "warning"},
        {"key": "ABIERTA_CRITICA", "label": "Críticas abiertas", "count": all_ordenes_qs.filter(prioridad__in=[OrdenMantenimiento.PRIORIDAD_CRITICA, OrdenMantenimiento.PRIORIDAD_ALTA], estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO]).count(), "tone": "danger"},
    ]
    enterprise_focus_cards = [
        {
            "label": "Órdenes críticas abiertas",
            "count": enterprise_cards[4]["count"],
            "tone": "danger",
            "detail": "Prioriza correctivos críticos o altos antes del cierre económico.",
            "url": f"{reverse('activos:ordenes')}?estatus=ABIERTAS&enterprise_gap=ABIERTA_CRITICA",
            "cta": "Atender críticas",
            "key": "ABIERTA_CRITICA",
        },
        {
            "label": "Órdenes sin responsable",
            "count": enterprise_cards[0]["count"],
            "tone": "warning",
            "detail": "Asigna dueño operativo para evitar órdenes huérfanas.",
            "url": f"{reverse('activos:ordenes')}?estatus={estado}&enterprise_gap=SIN_RESPONSABLE",
            "cta": "Asignar responsable",
            "key": "SIN_RESPONSABLE",
        },
        {
            "label": "Cierre económico pendiente",
            "count": enterprise_cards[2]["count"],
            "tone": "warning",
            "detail": "Las órdenes cerradas sin costo distorsionan control y presupuesto.",
            "url": f"{reverse('activos:ordenes')}?estatus={estado}&enterprise_gap=SIN_COSTO_CIERRE",
            "cta": "Completar costos",
            "key": "SIN_COSTO_CIERRE",
        },
    ]
    focus_summary = None
    selected_card = next((card for card in enterprise_cards if card["key"] == enterprise_gap), None)
    if selected_card:
        focus_summary = _active_focus_summary(
            focus_label=selected_card["label"],
            focus_detail=f"Vista filtrada por {selected_card['label'].lower()} para cierre documental.",
            clear_url=f"{reverse('activos:ordenes')}?estatus={estado}",
        )
    today = timezone.localdate()
    activos_total = Activo.objects.filter(activo=True).count()
    all_planes_qs = PlanMantenimiento.objects.filter(estatus=PlanMantenimiento.ESTATUS_ACTIVO, activo=True)
    active_plan_asset_ids = set(all_planes_qs.values_list("activo_ref_id", flat=True))
    overdue_plan_asset_ids = set(
        all_planes_qs.filter(proxima_ejecucion__isnull=False, proxima_ejecucion__lt=today).values_list("activo_ref_id", flat=True)
    )
    critical_open_asset_ids = set(
        all_ordenes_qs.filter(
            estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO],
            tipo=OrdenMantenimiento.TIPO_CORRECTIVO,
            prioridad__in=[OrdenMantenimiento.PRIORIDAD_CRITICA, OrdenMantenimiento.PRIORIDAD_ALTA],
        ).values_list("activo_ref_id", flat=True)
    )
    blockers_total = (
        Activo.objects.filter(Q(categoria__isnull=True) | Q(categoria=""), activo=True).count()
        + Activo.objects.filter(proveedor_mantenimiento__isnull=True, activo=True).count()
        + max(activos_total - len(active_plan_asset_ids), 0)
        + len(overdue_plan_asset_ids)
        + len(critical_open_asset_ids)
        + Activo.objects.filter(
            estado=Activo.ESTADO_FUERA_SERVICIO,
            criticidad=Activo.CRITICIDAD_ALTA,
            activo=True,
        ).count()
    )
    enterprise_chain = _activos_enterprise_chain(
        activos_total=activos_total,
        blockers_total=blockers_total,
        planes_total=all_planes_qs.count(),
        overdue_plans=len(overdue_plan_asset_ids),
        ordenes_abiertas=all_ordenes_qs.filter(
            estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO]
        ).count(),
        critical_open=len(critical_open_asset_ids),
        bitacora_30d=BitacoraMantenimiento.objects.filter(fecha__date__gte=today - timedelta(days=30)).count(),
    )
    enterprise_maturity_summary = _activos_maturity_summary(chain=enterprise_chain)
    enterprise_handoff_map = _activos_handoff_map(
        blockers_total=blockers_total,
        overdue_plans=len(overdue_plan_asset_ids),
        ordenes_abiertas=all_ordenes_qs.filter(
            estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO]
        ).count(),
        bitacora_30d=BitacoraMantenimiento.objects.filter(fecha__date__gte=today - timedelta(days=30)).count(),
    )
    context = {
        "module_tabs": _module_tabs("ordenes"),
        "ordenes_rows": ordenes_rows,
        "ordenes_editables": list(
            OrdenMantenimiento.objects.select_related("activo_ref")
            .exclude(estatus=OrdenMantenimiento.ESTATUS_CANCELADA)
            .order_by("-fecha_programada", "-id")[:600]
        ),
        "activos": list(Activo.objects.filter(activo=True).order_by("nombre")[:800]),
        "planes": list(
            PlanMantenimiento.objects.filter(
                estatus=PlanMantenimiento.ESTATUS_ACTIVO,
                activo=True,
                activo_ref__activo=True,
            )
            .select_related("activo_ref")
            .order_by("activo_ref__nombre", "nombre")[:1200]
        ),
        "estado": estado,
        "can_manage_activos": can_manage_inventario(request.user),
        "enterprise_gap": enterprise_gap,
        "enterprise_cards": enterprise_cards,
        "enterprise_focus_cards": enterprise_focus_cards,
        "focus_summary": focus_summary,
        "operational_health_cards": _activos_operational_health_cards(cards=enterprise_cards, module_key="ordenes"),
        "enterprise_chain": enterprise_chain,
        "enterprise_maturity_summary": enterprise_maturity_summary,
        "enterprise_handoff_map": enterprise_handoff_map,
        "document_stage_rows": _annotate_activos_document_stage_rows([
            {
                "label": "Órdenes abiertas",
                "open": all_ordenes_qs.filter(
                    estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO]
                ).count(),
                "closed": all_ordenes_qs.filter(estatus=OrdenMantenimiento.ESTATUS_CERRADA).count(),
                "detail": "Órdenes abiertas frente a órdenes documentadas como cerradas.",
                "url": reverse("activos:ordenes"),
            },
            {
                "label": "Control de ejecución",
                "open": enterprise_cards[0]["count"] + enterprise_cards[1]["count"],
                "closed": max(all_ordenes_qs.count() - (enterprise_cards[0]["count"] + enterprise_cards[1]["count"]), 0),
                "detail": "Órdenes sin responsable o sin fecha de inicio frente a órdenes trazables.",
                "url": reverse("activos:ordenes"),
            },
            {
                "label": "Cierre económico",
                "open": enterprise_cards[2]["count"],
                "closed": max(all_ordenes_qs.filter(estatus=OrdenMantenimiento.ESTATUS_CERRADA).count() - enterprise_cards[2]["count"], 0),
                "detail": "Cierres sin costo frente a cierres con evidencia económica.",
                "url": reverse("activos:ordenes"),
            },
            {
                "label": "Origen preventivo",
                "open": enterprise_cards[3]["count"],
                "closed": max(
                    all_ordenes_qs.filter(tipo=OrdenMantenimiento.TIPO_PREVENTIVO).count() - enterprise_cards[3]["count"],
                    0,
                ),
                "detail": "Órdenes preventivas ligadas a plan contra órdenes sueltas.",
                "url": reverse("activos:planes"),
            },
        ]),
        "erp_command_center": _activos_command_center(
            owner="Mantenimiento / Ejecución",
            blockers_total=sum(card["count"] for card in enterprise_cards),
            maturity_summary=enterprise_maturity_summary,
            default_url=reverse("activos:ordenes"),
            default_cta="Abrir órdenes",
        ),
    }
    context["erp_governance_rows"] = _activos_erp_governance_rows(context["document_stage_rows"])
    context["release_gate_rows"] = _activos_release_gate_rows(context["document_stage_rows"])
    context["release_gate_completion"] = _activos_release_gate_completion(context["release_gate_rows"])
    context["critical_path_rows"] = _activos_critical_path_rows(context["enterprise_chain"])
    context["executive_radar_rows"] = _activos_executive_radar_rows(
        context["document_stage_rows"],
        context["enterprise_chain"],
    )
    return render(request, "activos/ordenes.html", context)


@login_required
def actualizar_orden_estatus(request, pk: int, estatus: str):
    if request.method != "POST":
        return redirect("activos:ordenes")
    if not can_manage_inventario(request.user):
        raise PermissionDenied("No tienes permisos para gestionar órdenes de mantenimiento.")

    estatus = (estatus or "").strip().upper()
    if estatus not in {x[0] for x in OrdenMantenimiento.ESTATUS_CHOICES}:
        messages.error(request, "Estatus inválido.")
        return redirect("activos:ordenes")

    orden = get_object_or_404(OrdenMantenimiento, pk=pk)
    from_status = orden.estatus
    if from_status == estatus:
        return redirect("activos:ordenes")
    orden.estatus = estatus
    today = timezone.localdate()
    if estatus == OrdenMantenimiento.ESTATUS_EN_PROCESO and not orden.fecha_inicio:
        orden.fecha_inicio = today
    if estatus == OrdenMantenimiento.ESTATUS_CERRADA:
        orden.fecha_cierre = today
        if orden.plan_ref_id:
            plan = orden.plan_ref
            plan.ultima_ejecucion = today
            plan.recompute_next_date()
            plan.save(update_fields=["ultima_ejecucion", "proxima_ejecucion", "actualizado_en"])
    orden.save(update_fields=["estatus", "fecha_inicio", "fecha_cierre", "actualizado_en"])
    BitacoraMantenimiento.objects.create(
        orden=orden,
        accion="ESTATUS",
        comentario=f"{from_status} -> {estatus}",
        usuario=request.user,
    )
    log_event(
        request.user,
        "UPDATE",
        "activos.OrdenMantenimiento",
        orden.id,
        {"from": from_status, "to": estatus, "folio": orden.folio},
    )
    messages.success(request, f"Orden {orden.folio} actualizada a {estatus}.")
    return redirect("activos:ordenes")


@login_required
def reportes_servicio(request):
    if not can_view_inventario(request.user):
        raise PermissionDenied("No tienes permisos para ver Activos.")

    if request.method == "POST":
        activo_id = _safe_int(request.POST.get("activo_id"))
        descripcion = (request.POST.get("descripcion") or "").strip()
        if not activo_id or not descripcion:
            messages.error(request, "Activo y descripción del reporte son obligatorios.")
            return redirect("activos:reportes")
        activo_obj = get_object_or_404(Activo, pk=activo_id)
        prioridad = (request.POST.get("prioridad") or OrdenMantenimiento.PRIORIDAD_MEDIA).strip().upper()
        prioridad = (
            prioridad if prioridad in {x[0] for x in OrdenMantenimiento.PRIORIDAD_CHOICES} else OrdenMantenimiento.PRIORIDAD_MEDIA
        )
        perfil = getattr(request.user, "userprofile", None)
        area = perfil.departamento.nombre if perfil and perfil.departamento_id else ""
        sucursal = perfil.sucursal.nombre if perfil and perfil.sucursal_id else ""
        responsable = (request.POST.get("responsable") or "").strip() or request.user.get_full_name() or request.user.username
        fecha_programada = _parse_date(request.POST.get("fecha_programada")) or timezone.localdate()
        orden = OrdenMantenimiento.objects.create(
            activo_ref=activo_obj,
            tipo=OrdenMantenimiento.TIPO_CORRECTIVO,
            prioridad=prioridad,
            estatus=OrdenMantenimiento.ESTATUS_PENDIENTE,
            fecha_programada=fecha_programada,
            responsable=responsable,
            descripcion=descripcion,
            creado_por=request.user,
        )
        contexto = []
        if area:
            contexto.append(f"Área: {area}")
        if sucursal:
            contexto.append(f"Sucursal: {sucursal}")
        BitacoraMantenimiento.objects.create(
            orden=orden,
            accion="REPORTE_FALLA",
            comentario=" · ".join(contexto) if contexto else "Reporte desde módulo Activos",
            usuario=request.user,
        )
        log_event(
            request.user,
            "CREATE",
            "activos.OrdenMantenimiento",
            orden.id,
            {"folio": orden.folio, "tipo": "REPORTE_FALLA", "activo_id": activo_obj.id, "prioridad": prioridad},
        )
        messages.success(request, f"Reporte levantado. Orden generada: {orden.folio}.")
        return redirect("activos:reportes")

    estado = (request.GET.get("estatus") or "ABIERTAS").strip().upper()
    semaforo_filter = (request.GET.get("semaforo") or "").strip().upper()
    q = (request.GET.get("q") or "").strip()
    qs = OrdenMantenimiento.objects.select_related("activo_ref", "creado_por").filter(tipo=OrdenMantenimiento.TIPO_CORRECTIVO)
    if estado == "ABIERTAS":
        qs = qs.filter(estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO])
    elif estado in {x[0] for x in OrdenMantenimiento.ESTATUS_CHOICES}:
        qs = qs.filter(estatus=estado)
    if q:
        qs = qs.filter(Q(folio__icontains=q) | Q(activo_ref__nombre__icontains=q) | Q(descripcion__icontains=q))

    today = timezone.localdate()
    reportes = []
    for orden in qs.order_by("-fecha_programada", "-id")[:1000]:
        dias = (today - orden.fecha_programada).days if orden.fecha_programada else 0
        if orden.estatus == OrdenMantenimiento.ESTATUS_CERRADA:
            semaforo = ("Verde", "badge-success")
        elif dias <= 2:
            semaforo = ("Verde", "badge-success")
        elif dias <= 5:
            semaforo = ("Amarillo", "badge-warning")
        else:
            semaforo = ("Rojo", "badge-danger")
        reportes.append(
            {
                "orden": orden,
                "dias": dias,
                "semaforo_label": semaforo[0],
                "semaforo_class": semaforo[1],
                "semaforo_key": semaforo[0].upper(),
            }
        )

    if semaforo_filter in {"VERDE", "AMARILLO", "ROJO"}:
        reportes = [item for item in reportes if item.get("semaforo_key") == semaforo_filter]

    export_format = (request.GET.get("export") or "").strip().lower()
    if export_format in {"csv", "xlsx"}:
        if export_format == "csv":
            return _export_reportes_servicio_csv(reportes)
        return _export_reportes_servicio_xlsx(reportes)

    activos_total = Activo.objects.filter(activo=True).count()
    planes_total = PlanMantenimiento.objects.filter(estatus=PlanMantenimiento.ESTATUS_ACTIVO, activo=True).count()
    ordenes_abiertas_total = OrdenMantenimiento.objects.filter(
        estatus__in=[OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO]
    ).count()
    bitacora_30d_total = BitacoraMantenimiento.objects.filter(fecha__date__gte=today - timedelta(days=30)).count()
    blockers_total = len([item for item in reportes if item.get("semaforo_key") in {"ROJO", "AMARILLO"}])
    enterprise_chain = _activos_enterprise_chain(
        activos_total=activos_total,
        blockers_total=blockers_total,
        planes_total=planes_total,
        overdue_plans=0,
        ordenes_abiertas=ordenes_abiertas_total,
        critical_open=len([item for item in reportes if item.get("semaforo_key") == "ROJO"]),
        bitacora_30d=bitacora_30d_total,
    )
    enterprise_maturity_summary = _activos_maturity_summary(chain=enterprise_chain)
    enterprise_handoff_map = _activos_handoff_map(
        blockers_total=blockers_total,
        overdue_plans=0,
        ordenes_abiertas=ordenes_abiertas_total,
        bitacora_30d=bitacora_30d_total,
    )
    document_stage_rows = _annotate_activos_document_stage_rows([
        {
            "label": "Reportes correctivos",
            "open": len(reportes),
            "closed": max(bitacora_30d_total - len(reportes), 0),
            "detail": "Reportes abiertos vs evidencia registrada en bitácora reciente.",
            "url": reverse("activos:reportes"),
        },
        {
            "label": "Órdenes abiertas",
            "open": ordenes_abiertas_total,
            "closed": bitacora_30d_total,
            "detail": "Seguimiento a correctivos pendientes y atendidos.",
            "url": reverse("activos:ordenes"),
        },
        {
            "label": "Planes preventivos",
            "open": 0,
            "closed": planes_total,
            "detail": "Base preventiva disponible para priorizar reportes.",
            "url": reverse("activos:planes"),
        },
    ])
    report_focus_cards = [
        {
            "label": "Atención inmediata",
            "count": len([item for item in reportes if item.get("semaforo_key") == "ROJO"]),
            "tone": "danger",
            "detail": "Reportes correctivos vencidos o críticos que ya deben entrar a ejecución.",
            "url": f"{reverse('activos:reportes')}?estatus={estado}&semaforo=ROJO&q={q}",
            "cta": "Enfocar rojos",
            "key": "ROJO",
        },
        {
            "label": "Seguimiento operativo",
            "count": len([item for item in reportes if item.get("semaforo_key") == "AMARILLO"]),
            "tone": "warning",
            "detail": "Incidentes que siguen abiertos y requieren confirmación de arranque.",
            "url": f"{reverse('activos:reportes')}?estatus={estado}&semaforo=AMARILLO&q={q}",
            "cta": "Enfocar amarillos",
            "key": "AMARILLO",
        },
        {
            "label": "Documentados al día",
            "count": len([item for item in reportes if item.get("semaforo_key") == "VERDE"]),
            "tone": "success",
            "detail": "Incidentes controlados o cerrados dentro del SLA operativo.",
            "url": f"{reverse('activos:reportes')}?estatus={estado}&semaforo=VERDE&q={q}",
            "cta": "Ver verdes",
            "key": "VERDE",
        },
    ]
    report_focus_summary = None
    if semaforo_filter in {"VERDE", "AMARILLO", "ROJO"}:
        report_focus_summary = _active_focus_summary(
            focus_label=f"Semáforo {semaforo_filter.title()}",
            focus_detail="Vista enfocada al subconjunto documental seleccionado para seguimiento o cierre.",
            clear_url=f"{reverse('activos:reportes')}?estatus={estado}&q={q}",
        )

    context = {
        "module_tabs": _module_tabs("reportes"),
        "activos": list(Activo.objects.filter(activo=True).order_by("nombre")[:800]),
        "reportes": reportes,
        "estado": estado,
        "semaforo_filter": semaforo_filter,
        "q": q,
        "today": today,
        "prioridad_choices": OrdenMantenimiento.PRIORIDAD_CHOICES,
        "can_manage_activos": can_manage_inventario(request.user),
        "operational_health_cards": _activos_operational_health_cards(
            cards=[
                {
                    "count": len([item for item in reportes if item.get("semaforo_key") == "AMARILLO"]),
                    "tone": "warning",
                },
                {
                    "count": len([item for item in reportes if item.get("semaforo_key") == "ROJO"]),
                    "tone": "danger",
                },
            ],
            module_key="reportes",
        ),
        "enterprise_chain": enterprise_chain,
        "enterprise_maturity_summary": enterprise_maturity_summary,
        "enterprise_handoff_map": enterprise_handoff_map,
        "document_stage_rows": document_stage_rows,
        "erp_governance_rows": _activos_erp_governance_rows(document_stage_rows),
        "enterprise_focus_cards": report_focus_cards,
        "focus_summary": report_focus_summary,
        "erp_command_center": _activos_command_center(
            owner="Mantenimiento / Ejecución",
            blockers_total=blockers_total,
            maturity_summary=enterprise_maturity_summary,
            default_url=reverse("activos:reportes"),
            default_cta="Abrir reportes",
        ),
    }
    context["release_gate_rows"] = _activos_release_gate_rows(context["document_stage_rows"])
    context["release_gate_completion"] = _activos_release_gate_completion(context["release_gate_rows"])
    context["critical_path_rows"] = _activos_critical_path_rows(context["enterprise_chain"])
    context["executive_radar_rows"] = _activos_executive_radar_rows(
        context["document_stage_rows"],
        context["enterprise_chain"],
    )
    return render(request, "activos/reportes.html", context)


@login_required
def calendario(request):
    if not can_view_inventario(request.user):
        raise PermissionDenied("No tienes permisos para ver Activos.")

    date_from_raw = (request.GET.get("from") or "").strip()
    date_to_raw = (request.GET.get("to") or "").strip()
    days_param = _safe_int(request.GET.get("days"), default=45)
    if days_param <= 0:
        days_param = 45
    days_param = max(7, min(120, days_param))
    try:
        date_from = timezone.datetime.fromisoformat(date_from_raw).date() if date_from_raw else timezone.localdate()
    except ValueError:
        date_from = timezone.localdate()
    try:
        date_to = (
            timezone.datetime.fromisoformat(date_to_raw).date()
            if date_to_raw
            else (date_from + timedelta(days=days_param))
        )
    except ValueError:
        date_to = date_from + timedelta(days=days_param)
    if date_to < date_from:
        date_to = date_from + timedelta(days=days_param)

    planes = list(
        PlanMantenimiento.objects.select_related("activo_ref")
        .filter(
            activo=True,
            estatus=PlanMantenimiento.ESTATUS_ACTIVO,
            proxima_ejecucion__isnull=False,
            proxima_ejecucion__gte=date_from,
            proxima_ejecucion__lte=date_to,
        )
        .order_by("proxima_ejecucion", "id")
    )
    ordenes_qs = (
        OrdenMantenimiento.objects.select_related("activo_ref", "plan_ref")
        .filter(fecha_programada__gte=date_from, fecha_programada__lte=date_to)
        .order_by("fecha_programada", "id")
    )

    events = []
    for plan in planes:
        events.append(
            {
                "fecha": plan.proxima_ejecucion,
                "tipo": "Plan",
                "referencia": f"Plan #{plan.id}",
                "activo": plan.activo_ref.nombre,
                "detalle": plan.nombre,
                "estado": plan.estatus,
            }
        )
    for orden in ordenes_qs:
        events.append(
            {
                "fecha": orden.fecha_programada,
                "tipo": "Orden",
                "referencia": orden.folio,
                "activo": orden.activo_ref.nombre,
                "detalle": orden.descripcion or orden.get_tipo_display(),
                "estado": orden.estatus,
            }
        )
    events.sort(key=lambda r: (r["fecha"], r["tipo"], r["referencia"]))

    ordenes_list = list(ordenes_qs)
    resumen = {
        "planes": len(planes),
        "ordenes_total": len(ordenes_list),
        "ordenes_pendientes": sum(1 for o in ordenes_list if o.estatus == OrdenMantenimiento.ESTATUS_PENDIENTE),
        "ordenes_en_proceso": sum(1 for o in ordenes_list if o.estatus == OrdenMantenimiento.ESTATUS_EN_PROCESO),
        "ordenes_cerradas": sum(1 for o in ordenes_list if o.estatus == OrdenMantenimiento.ESTATUS_CERRADA),
    }
    enterprise_chain = _activos_enterprise_chain(
        activos_total=Activo.objects.filter(activo=True).count(),
        blockers_total=resumen["ordenes_pendientes"] + resumen["ordenes_en_proceso"],
        planes_total=PlanMantenimiento.objects.filter(estatus=PlanMantenimiento.ESTATUS_ACTIVO, activo=True).count(),
        overdue_plans=0,
        ordenes_abiertas=sum(
            1
            for o in ordenes_list
            if o.estatus in [OrdenMantenimiento.ESTATUS_PENDIENTE, OrdenMantenimiento.ESTATUS_EN_PROCESO]
        ),
        critical_open=resumen["ordenes_en_proceso"],
        bitacora_30d=BitacoraMantenimiento.objects.filter(fecha__date__gte=timezone.localdate() - timedelta(days=30)).count(),
    )
    enterprise_maturity_summary = _activos_maturity_summary(chain=enterprise_chain)
    enterprise_handoff_map = _activos_handoff_map(
        blockers_total=resumen["ordenes_pendientes"] + resumen["ordenes_en_proceso"],
        overdue_plans=0,
        ordenes_abiertas=resumen["ordenes_pendientes"] + resumen["ordenes_en_proceso"],
        bitacora_30d=BitacoraMantenimiento.objects.filter(
            fecha__date__gte=timezone.localdate() - timedelta(days=30)
        ).count(),
    )
    document_stage_rows = _annotate_activos_document_stage_rows([
        {
            "label": "Planes en ventana",
            "open": resumen["planes"],
            "closed": 0,
            "detail": "Eventos preventivos programados dentro del rango consultado.",
            "url": reverse("activos:planes"),
        },
        {
            "label": "Órdenes en ventana",
            "open": resumen["ordenes_pendientes"] + resumen["ordenes_en_proceso"],
            "closed": resumen["ordenes_cerradas"],
            "detail": "Seguimiento documental de órdenes dentro del periodo.",
            "url": reverse("activos:ordenes"),
        },
        {
            "label": "Bitácora reciente",
            "open": 0,
            "closed": BitacoraMantenimiento.objects.filter(
                fecha__date__gte=timezone.localdate() - timedelta(days=30)
            ).count(),
            "detail": "Eventos de mantenimiento registrados en últimos 30 días.",
            "url": reverse("activos:reportes"),
        },
    ])

    context = {
        "module_tabs": _module_tabs("calendario"),
        "date_from": date_from,
        "date_to": date_to,
        "days": days_param,
        "events": events,
        "resumen": resumen,
        "operational_health_cards": _activos_operational_health_cards(
            cards=[
                {"count": resumen["ordenes_pendientes"], "tone": "warning"},
                {"count": resumen["ordenes_en_proceso"], "tone": "danger"},
            ],
            module_key="calendario",
        ),
        "enterprise_chain": enterprise_chain,
        "enterprise_maturity_summary": enterprise_maturity_summary,
        "enterprise_handoff_map": enterprise_handoff_map,
        "document_stage_rows": document_stage_rows,
        "erp_governance_rows": _activos_erp_governance_rows(document_stage_rows),
        "erp_command_center": _activos_command_center(
            owner="Planeación / Operación",
            blockers_total=resumen["ordenes_pendientes"] + resumen["ordenes_en_proceso"],
            maturity_summary=enterprise_maturity_summary,
            default_url=reverse("activos:calendario"),
            default_cta="Abrir calendario",
        ),
    }
    context["release_gate_rows"] = _activos_release_gate_rows(context["document_stage_rows"])
    context["release_gate_completion"] = _activos_release_gate_completion(context["release_gate_rows"])
    context["critical_path_rows"] = _activos_critical_path_rows(context["enterprise_chain"])
    context["executive_radar_rows"] = _activos_executive_radar_rows(
        context["document_stage_rows"],
        context["enterprise_chain"],
    )
    return render(request, "activos/calendario.html", context)
