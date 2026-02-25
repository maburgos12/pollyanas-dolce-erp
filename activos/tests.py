from io import BytesIO
from datetime import timedelta

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from core.access import ROLE_ADMIN, ROLE_ALMACEN, ROLE_VENTAS
from core.models import AuditLog

from .models import Activo, OrdenMantenimiento, PlanMantenimiento
from django.utils import timezone


class ActivosFlowsTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.admin = user_model.objects.create_user("admin_activos", "admin_activos@example.com", "test12345")
        self.almacen = user_model.objects.create_user("almacen_activos", "almacen_activos@example.com", "test12345")
        self.ventas = user_model.objects.create_user("ventas_activos", "ventas_activos@example.com", "test12345")

        Group.objects.get_or_create(name=ROLE_ADMIN)[0].user_set.add(self.admin)
        Group.objects.get_or_create(name=ROLE_ALMACEN)[0].user_set.add(self.almacen)
        Group.objects.get_or_create(name=ROLE_VENTAS)[0].user_set.add(self.ventas)

    def test_admin_can_create_activo_from_ui(self):
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("activos:activos"),
            {
                "action": "create_activo",
                "nombre": "Refrigerador Cámara 01",
                "categoria": "Refrigeración",
                "estado": "OPERATIVO",
                "criticidad": "ALTA",
                "activo": "1",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(Activo.objects.filter(nombre="Refrigerador Cámara 01").exists())

    def test_almacen_can_raise_service_report(self):
        activo = Activo.objects.create(nombre="AA Oficina", categoria="Aire")
        self.client.force_login(self.almacen)
        response = self.client.post(
            reverse("activos:reportes"),
            {
                "activo_id": str(activo.id),
                "prioridad": "MEDIA",
                "descripcion": "No enfría correctamente",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            OrdenMantenimiento.objects.filter(activo_ref=activo, tipo=OrdenMantenimiento.TIPO_CORRECTIVO).exists()
        )

    def test_ventas_cannot_access_activos_module(self):
        self.client.force_login(self.ventas)
        response = self.client.get(reverse("activos:activos"))
        self.assertEqual(response.status_code, 403)

    def test_admin_can_create_plan(self):
        activo = Activo.objects.create(nombre="Horno 02", categoria="Hornos")
        self.client.force_login(self.admin)
        response = self.client.post(
            reverse("activos:planes"),
            {
                "action": "create_plan",
                "activo_id": str(activo.id),
                "nombre": "Mantenimiento mensual horno",
                "tipo": PlanMantenimiento.TIPO_PREVENTIVO,
                "frecuencia_dias": "30",
                "estatus": PlanMantenimiento.ESTATUS_ACTIVO,
                "activo": "1",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(PlanMantenimiento.objects.filter(activo_ref=activo).exists())

    def test_export_planes_csv(self):
        self.client.force_login(self.admin)
        activo = Activo.objects.create(nombre="Horno QA", categoria="Hornos")
        PlanMantenimiento.objects.create(
            activo_ref=activo,
            nombre="Plan QA",
            estatus=PlanMantenimiento.ESTATUS_ACTIVO,
            activo=True,
            frecuencia_dias=30,
            proxima_ejecucion=timezone.localdate(),
        )
        response = self.client.get(reverse("activos:planes"), {"export": "csv"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response.get("Content-Type", ""))
        self.assertIn("activos_planes_", response.get("Content-Disposition", ""))
        body = response.content.decode("utf-8")
        self.assertIn("activo_codigo,activo,plan,tipo,estatus", body)
        self.assertIn("Plan QA", body)

    def test_export_ordenes_xlsx(self):
        self.client.force_login(self.admin)
        activo = Activo.objects.create(nombre="Batidora QA", categoria="Batidoras")
        OrdenMantenimiento.objects.create(
            activo_ref=activo,
            tipo=OrdenMantenimiento.TIPO_PREVENTIVO,
            estatus=OrdenMantenimiento.ESTATUS_PENDIENTE,
            fecha_programada=timezone.localdate(),
            descripcion="Orden QA",
        )
        response = self.client.get(reverse("activos:ordenes"), {"export": "xlsx", "estatus": "ABIERTAS"})
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.get("Content-Type", ""),
        )
        self.assertIn("activos_ordenes_", response.get("Content-Disposition", ""))
        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(
            headers[:7],
            ["folio", "activo_codigo", "activo", "plan", "tipo", "prioridad", "estatus"],
        )

    def test_admin_can_update_orden_costos_and_close(self):
        self.client.force_login(self.admin)
        activo = Activo.objects.create(nombre="Conservador QA", categoria="Refrigeración")
        orden = OrdenMantenimiento.objects.create(
            activo_ref=activo,
            tipo=OrdenMantenimiento.TIPO_CORRECTIVO,
            estatus=OrdenMantenimiento.ESTATUS_EN_PROCESO,
            fecha_programada=timezone.localdate(),
            descripcion="Servicio costo",
        )
        response = self.client.post(
            reverse("activos:ordenes"),
            {
                "action": "update_costos",
                "orden_id": str(orden.id),
                "costo_repuestos": "1500.25",
                "costo_mano_obra": "800",
                "costo_otros": "120",
                "cerrar_orden": "1",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        orden.refresh_from_db()
        self.assertEqual(str(orden.costo_repuestos), "1500.25")
        self.assertEqual(str(orden.costo_mano_obra), "800.00")
        self.assertEqual(str(orden.costo_otros), "120.00")
        self.assertEqual(orden.estatus, OrdenMantenimiento.ESTATUS_CERRADA)

    def test_calendario_days_window(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("activos:calendario"), {"days": "15"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["days"], 15)

    def test_export_reportes_servicio_csv(self):
        self.client.force_login(self.admin)
        activo = Activo.objects.create(nombre="Refrigerador QA", categoria="Refrigeración")
        OrdenMantenimiento.objects.create(
            activo_ref=activo,
            tipo=OrdenMantenimiento.TIPO_CORRECTIVO,
            prioridad=OrdenMantenimiento.PRIORIDAD_MEDIA,
            estatus=OrdenMantenimiento.ESTATUS_PENDIENTE,
            fecha_programada=timezone.localdate(),
            descripcion="Falla de prueba",
        )
        response = self.client.get(
            reverse("activos:reportes"),
            {"export": "csv", "estatus": "ABIERTAS"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response.get("Content-Type", ""))
        self.assertIn("activos_reportes_servicio_", response.get("Content-Disposition", ""))
        body = response.content.decode("utf-8")
        self.assertIn("folio,fecha,activo_codigo,activo,prioridad,estatus,semaforo,dias,descripcion,responsable", body)
        self.assertIn("Falla de prueba", body)

    def test_filter_reportes_servicio_by_semaforo(self):
        self.client.force_login(self.admin)
        activo = Activo.objects.create(nombre="Horno QA2", categoria="Hornos")
        OrdenMantenimiento.objects.create(
            activo_ref=activo,
            tipo=OrdenMantenimiento.TIPO_CORRECTIVO,
            prioridad=OrdenMantenimiento.PRIORIDAD_MEDIA,
            estatus=OrdenMantenimiento.ESTATUS_PENDIENTE,
            fecha_programada=timezone.localdate() - timedelta(days=8),
            descripcion="Falla roja",
        )
        response = self.client.get(
            reverse("activos:reportes"),
            {"estatus": "ABIERTAS", "semaforo": "ROJO"},
        )
        self.assertEqual(response.status_code, 200)
        reportes = response.context["reportes"]
        self.assertTrue(reportes)
        self.assertTrue(all(item.get("semaforo_key") == "ROJO" for item in reportes))

    def test_dashboard_alertas_criticas_context(self):
        self.client.force_login(self.admin)
        activo = Activo.objects.create(nombre="AA Critico", categoria="Aire", criticidad=Activo.CRITICIDAD_ALTA)
        PlanMantenimiento.objects.create(
            activo_ref=activo,
            nombre="Plan vencido QA",
            estatus=PlanMantenimiento.ESTATUS_ACTIVO,
            activo=True,
            proxima_ejecucion=timezone.localdate() - timedelta(days=3),
            frecuencia_dias=30,
        )
        OrdenMantenimiento.objects.create(
            activo_ref=activo,
            tipo=OrdenMantenimiento.TIPO_CORRECTIVO,
            prioridad=OrdenMantenimiento.PRIORIDAD_CRITICA,
            estatus=OrdenMantenimiento.ESTATUS_PENDIENTE,
            fecha_programada=timezone.localdate(),
            descripcion="Orden critica QA",
        )
        response = self.client.get(reverse("activos:dashboard"))
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["planes_vencidos_rows"])
        self.assertTrue(response.context["ordenes_criticas_rows"])

    def test_generar_ordenes_programadas_creates_preventive_order(self):
        self.client.force_login(self.admin)
        activo = Activo.objects.create(nombre="AA Planta", categoria="Aire", criticidad=Activo.CRITICIDAD_ALTA)
        plan = PlanMantenimiento.objects.create(
            activo_ref=activo,
            nombre="Plan semanal",
            estatus=PlanMantenimiento.ESTATUS_ACTIVO,
            activo=True,
            proxima_ejecucion=timezone.localdate(),
            frecuencia_dias=7,
        )
        response = self.client.post(
            reverse("activos:planes"),
            {
                "action": "generar_ordenes_programadas",
                "scope": "overdue",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            OrdenMantenimiento.objects.filter(
                plan_ref=plan,
                tipo=OrdenMantenimiento.TIPO_PREVENTIVO,
                fecha_programada=timezone.localdate(),
            ).exists()
        )

    def test_generar_ordenes_programadas_avoids_duplicates(self):
        self.client.force_login(self.admin)
        today = timezone.localdate()
        activo = Activo.objects.create(nombre="Horno Línea 2", categoria="Hornos")
        plan = PlanMantenimiento.objects.create(
            activo_ref=activo,
            nombre="Plan mensual",
            estatus=PlanMantenimiento.ESTATUS_ACTIVO,
            activo=True,
            proxima_ejecucion=today,
            frecuencia_dias=30,
        )
        OrdenMantenimiento.objects.create(
            activo_ref=activo,
            plan_ref=plan,
            tipo=OrdenMantenimiento.TIPO_PREVENTIVO,
            estatus=OrdenMantenimiento.ESTATUS_PENDIENTE,
            fecha_programada=today,
            descripcion="Existente",
        )
        response = self.client.post(
            reverse("activos:planes"),
            {
                "action": "generar_ordenes_programadas",
                "scope": "overdue",
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            OrdenMantenimiento.objects.filter(plan_ref=plan, fecha_programada=today).count(),
            1,
        )

    def test_export_activos_depuracion_csv(self):
        self.client.force_login(self.admin)
        Activo.objects.create(nombre="MATRIZ", categoria="Equipos", notas="")
        response = self.client.get(reverse("activos:activos"), {"export": "depuracion_csv"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response.get("Content-Type", ""))
        self.assertIn("activos_pendientes_depuracion_", response.get("Content-Disposition", ""))
        body = response.content.decode("utf-8")
        self.assertIn("codigo,nombre,ubicacion,categoria,estado,notas,motivos,acciones_sugeridas", body)
        self.assertIn("MATRIZ", body)

    def test_export_activos_depuracion_xlsx(self):
        self.client.force_login(self.admin)
        Activo.objects.create(nombre="NIO", categoria="Equipos", notas="")
        response = self.client.get(reverse("activos:activos"), {"export": "depuracion_xlsx"})
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.get("Content-Type", ""),
        )
        self.assertIn("activos_pendientes_depuracion_", response.get("Content-Disposition", ""))
        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(
            headers,
            ["codigo", "nombre", "ubicacion", "categoria", "estado", "notas", "motivos", "acciones_sugeridas"],
        )

    def test_export_template_bitacora_csv(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("activos:activos"), {"export": "template_bitacora_csv"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response.get("Content-Type", ""))
        self.assertIn("plantilla_bitacora_activos.csv", response.get("Content-Disposition", ""))
        body = response.content.decode("utf-8")
        self.assertIn("nombre,marca,modelo,serie,fecha_1,costo_1,fecha_2,costo_2", body)

    def test_export_template_bitacora_xlsx(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse("activos:activos"), {"export": "template_bitacora_xlsx"})
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.get("Content-Type", ""),
        )
        self.assertIn("plantilla_bitacora_activos.xlsx", response.get("Content-Disposition", ""))
        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, ["nombre", "marca", "modelo", "serie", "fecha_1", "costo_1", "fecha_2", "costo_2"])

    def test_admin_can_import_bitacora_from_ui_dry_run(self):
        self.client.force_login(self.admin)
        upload = self._build_bitacora_upload("bitacora_dryrun.xlsx")
        response = self.client.post(
            reverse("activos:activos"),
            {
                "action": "import_bitacora",
                "dry_run": "1",
                "archivo_bitacora": upload,
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Activo.objects.filter(nombre="HORNO TEST UI").exists())

    def test_admin_can_import_bitacora_from_ui_apply(self):
        self.client.force_login(self.admin)
        upload = self._build_bitacora_upload("bitacora_apply.xlsx")
        response = self.client.post(
            reverse("activos:activos"),
            {
                "action": "import_bitacora",
                "archivo_bitacora": upload,
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(Activo.objects.filter(nombre="HORNO TEST UI").exists())
        self.assertTrue(OrdenMantenimiento.objects.filter(descripcion__icontains="bitácora histórica").exists())
        self.assertTrue(
            AuditLog.objects.filter(action="IMPORT", model="activos.BitacoraImport", user=self.admin).exists()
        )

    def test_activos_view_shows_import_runs_block(self):
        AuditLog.objects.create(
            user=self.admin,
            action="IMPORT",
            model="activos.BitacoraImport",
            object_id="demo",
            payload={"filename": "bitacora_test.csv", "filas_validas": 5, "source_format": "CSV"},
        )
        self.client.force_login(self.admin)
        response = self.client.get(reverse("activos:activos"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Últimas importaciones de bitácora")
        self.assertContains(response, "bitacora_test.csv")

    def test_admin_can_import_bitacora_csv_from_ui_apply(self):
        self.client.force_login(self.admin)
        upload = self._build_bitacora_csv_upload("bitacora_apply.csv")
        response = self.client.post(
            reverse("activos:activos"),
            {
                "action": "import_bitacora",
                "archivo_bitacora": upload,
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(Activo.objects.filter(nombre="HORNO TEST CSV").exists())
        self.assertTrue(OrdenMantenimiento.objects.filter(descripcion__icontains="bitácora histórica").exists())

    def test_admin_can_import_bitacora_csv_semicolon_decimal_comma(self):
        self.client.force_login(self.admin)
        upload = self._build_bitacora_csv_upload("bitacora_decimal.csv", semicolon_decimal=True)
        response = self.client.post(
            reverse("activos:activos"),
            {
                "action": "import_bitacora",
                "archivo_bitacora": upload,
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        activo = Activo.objects.get(nombre="HORNO TEST CSV DECIMAL")
        orden = OrdenMantenimiento.objects.filter(activo_ref=activo).order_by("-id").first()
        self.assertIsNotNone(orden)
        self.assertEqual(str(orden.costo_otros), "1250.75")

    @staticmethod
    def _build_bitacora_upload(filename: str) -> SimpleUploadedFile:
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1"
        ws.cell(2, 2, "HORNOS")
        ws.cell(2, 3, "MARCA")
        ws.cell(2, 4, "MODELO")
        ws.cell(2, 5, "SERIE:")
        ws.cell(2, 6, "FECHA MANTENIMIENTO")
        ws.cell(2, 7, "COSTO")
        ws.cell(2, 8, "FECHA MANTENIMIENTO")
        ws.cell(2, 9, "COSTO")
        ws.cell(3, 2, "PRODUCCION MATRIZ")
        ws.cell(4, 2, "HORNO TEST UI")
        ws.cell(4, 3, "ALPHA")
        ws.cell(4, 4, "HX-10")
        ws.cell(4, 5, "SER-001")
        ws.cell(4, 6, "2026-02-20")
        ws.cell(4, 7, 1200)
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return SimpleUploadedFile(
            filename,
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @staticmethod
    def _build_bitacora_csv_upload(filename: str, *, semicolon_decimal: bool = False) -> SimpleUploadedFile:
        if semicolon_decimal:
            csv_content = "\n".join(
                [
                    "nombre;marca;modelo;serie;fecha_1;costo_1;fecha_2;costo_2",
                    "HORNO TEST CSV DECIMAL;ALPHA;HX-12;SER-003;2026-02-20;1.250,75;;",
                ]
            )
        else:
            csv_content = "\n".join(
                [
                    "nombre,marca,modelo,serie,fecha_1,costo_1,fecha_2,costo_2",
                    "HORNO TEST CSV,ALPHA,HX-11,SER-002,2026-02-20,950.5,,",
                ]
            )
        return SimpleUploadedFile(
            filename,
            csv_content.encode("utf-8"),
            content_type="text/csv",
        )
