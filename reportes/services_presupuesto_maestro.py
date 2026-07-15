from __future__ import annotations

import csv
import unicodedata
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from django.db import transaction
from django.db.models import Q, Sum
from django.utils.text import slugify
from openpyxl import load_workbook

from core.branch_catalog import indice_sucursales_por_texto, resolver_sucursal_por_texto
from core.models import Sucursal
from reportes.models import (
    AreaPresupuesto,
    EmpresaResultadoMensual,
    LineaPresupuestoMensual,
    RubroPresupuesto,
)
from reportes.services_budget_vs_actual import MONTH_COLUMNS, money_decimal, parse_period, variance_pct


MONTH_NAME_TO_NUMBER = {month_name: month_number for month_name, month_number in MONTH_COLUMNS}

AREA_DEFINITIONS = [
    ("ventas", "Ventas", 10),
    ("produccion", "Producción", 20),
    ("gastos-venta", "Gastos de Venta", 30),
    ("administracion", "Administración", 40),
    ("nomina", "Nómina", 50),
    ("logistica", "Logística", 60),
    ("compras", "Compras / Costo", 70),
    ("capex", "CAPEX apertura", 80),
]

RUBRO_TYPES = {
    RubroPresupuesto.TIPO_INGRESO,
    RubroPresupuesto.TIPO_EGRESO,
    RubroPresupuesto.TIPO_COSTO,
    RubroPresupuesto.TIPO_CAPEX,
}

ACTUAL_ALIAS = {
    "ventas": "ventas",
    "venta": "ventas",
    "ingresos": "ventas",
    "costo_mp": "costo_mp",
    "costo materia prima": "costo_mp",
    "materia prima": "costo_mp",
    "compras": "costo_mp",
    "costo_reventa": "costo_reventa",
    "reventa": "costo_reventa",
    "gasto_fijo": "gasto_fijo",
    "gastos venta": "gasto_fijo",
    "gastos de venta": "gasto_fijo",
    "administracion": "gasto_fijo",
    "administración": "gasto_fijo",
    "mano_obra": "mano_obra",
    "mano de obra": "mano_obra",
    "nomina": "mano_obra",
    "nómina": "mano_obra",
    "indirectos": "indirectos",
    "produccion": "indirectos",
    "producción": "indirectos",
    "utilidad_operativa": "utilidad_operativa",
    "utilidad operativa": "utilidad_operativa",
}

ACTUAL_LABELS = {
    "ventas": "Venta total",
    "costo_mp": "Costo materia prima",
    "costo_reventa": "Costo reventa",
    "gasto_fijo": "Gasto fijo",
    "mano_obra": "Mano de obra producción",
    "indirectos": "Indirectos producción",
    "utilidad_operativa": "Utilidad operativa",
}

CAPEX_GUAMUCHIL_ROWS = [
    ("CAPEX Guamúchil local", 2026, 1, Decimal("101518.61")),
    ("CAPEX Guamúchil arquitecta", 2026, 1, Decimal("102300.00")),
    ("CAPEX Guamúchil local", 2026, 2, Decimal("83405.71")),
    ("CAPEX Guamúchil arquitecta", 2026, 2, Decimal("110000.00")),
    ("CAPEX Guamúchil local", 2026, 3, Decimal("294624.92")),
    ("CAPEX Guamúchil arquitecta", 2026, 3, Decimal("125000.00")),
    ("CAPEX Guamúchil equipo", 2026, 2, Decimal("159806.30")),
    ("CAPEX Guamúchil equipo", 2026, 3, Decimal("159806.29")),
]

TOTALIZER_KEYWORDS = (
    "total",
    "subtotal",
    "utilidad",
    "perdida",
    "pérdida",
    "presupuesto",
)

SECTION_HEADER_CONCEPTS = {
    "produccion",
    "producción",
    "logistica",
    "logística",
    "ingresos",
    "egresos",
    "ingreso",
    "egreso",
    "costos",
    "gastos",
    "administracion",
    "administración",
    "nomina",
    "nómina",
}

SALES_PARENT_CONCEPTS = {
    "bollo",
    "pastel mini",
    "pastel chico",
    "pastel mediano",
    "pastel grande",
    "pastel rebanadas",
    "rosca",
    "pay mediano",
    "pay grande",
    "pay rebanada",
    "flan",
    "pan de la casa",
    "galleta",
    "cheesecake",
    "vaso preparado",
    "bebidas/otros",
}


@dataclass(frozen=True)
class PresupuestoImportSummary:
    area: str
    version: str
    year: int
    rubros_created: int
    rubros_updated: int
    lines_created: int
    lines_updated: int
    skipped_rows: int


@dataclass(frozen=True)
class VentasReimportSummary:
    year: int
    version: str
    dry_run: bool
    deleted_rubros: int
    deleted_lines: int
    rubros_created: int
    rubros_updated: int
    lines_created: int
    lines_updated: int
    skipped_rows: int
    monthly_totals: dict[str, Decimal]
    unidades_upsertadas: int = 0
    unidades_sin_receta: tuple[str, ...] = ()


def normalize_area_code(value: str) -> str:
    raw = slugify(str(value or "").strip()).replace("_", "-")
    aliases = {
        "gastos-ventas": "gastos-venta",
        "gasto-venta": "gastos-venta",
        "compras-costo": "compras",
        "costos": "compras",
        "produccion": "produccion",
        "produccion-": "produccion",
    }
    return aliases.get(raw, raw)


def normalize_version(value: str | None) -> str:
    raw = str(value or "").strip().upper()
    if raw in {LineaPresupuestoMensual.VERSION_ORIGINAL, LineaPresupuestoMensual.VERSION_REVISADO}:
        return raw
    return LineaPresupuestoMensual.VERSION_ORIGINAL


def normalize_rubro_type(value: str | None, area_code: str) -> str:
    raw = str(value or "").strip().upper()
    if raw in RUBRO_TYPES:
        return raw
    if area_code == "ventas":
        return RubroPresupuesto.TIPO_INGRESO
    if area_code == "capex":
        return RubroPresupuesto.TIPO_CAPEX
    if area_code in {"compras", "produccion"}:
        return RubroPresupuesto.TIPO_COSTO
    return RubroPresupuesto.TIPO_EGRESO


def infer_rubro_type(value: str | None, area_code: str, concept: str) -> str:
    raw = str(value or "").strip().upper()
    if raw in RUBRO_TYPES:
        return raw
    normalized_concept = normalize_concept_text(concept)
    if area_code == "ventas":
        return RubroPresupuesto.TIPO_INGRESO
    if area_code == "capex":
        return RubroPresupuesto.TIPO_CAPEX
    if area_code == "compras":
        return RubroPresupuesto.TIPO_COSTO
    if area_code == "produccion":
        if "costo" in normalized_concept or "insumo" in normalized_concept or "producto" in normalized_concept:
            return RubroPresupuesto.TIPO_COSTO
        return RubroPresupuesto.TIPO_EGRESO
    return RubroPresupuesto.TIPO_EGRESO


def normalize_actual_key(*values: str) -> str:
    joined = " ".join(str(value or "") for value in values).strip().lower()
    compact = " ".join(joined.replace("-", " ").replace("_", " ").split())
    if compact in ACTUAL_ALIAS:
        return ACTUAL_ALIAS[compact]
    slug = slugify(compact).replace("-", "_")
    return ACTUAL_ALIAS.get(slug, "")


def money(value) -> Decimal:
    return money_decimal(value)


def normalize_concept_text(value: object) -> str:
    return " ".join(str(value or "").strip().lower().split())


def normalize_header_text(value: object) -> str:
    text = normalize_concept_text(value)
    text = "".join(
        char for char in unicodedata.normalize("NFKD", text) if not unicodedata.combining(char)
    )
    return text.replace("_", " ").replace("-", " ")


def is_totalizer_budget_concept(concept: object, *, area_code: str = "") -> bool:
    normalized = normalize_concept_text(concept)
    if not normalized:
        return True
    if normalized in {"gasto", "cuenta", "concepto", "conceptos", "descripcion", "descripción"}:
        return True
    if any(keyword in normalized for keyword in TOTALIZER_KEYWORDS):
        return True
    if normalized in SECTION_HEADER_CONCEPTS:
        return True
    normalized_area = normalize_concept_text(area_code).replace("-", " ")
    if normalized_area and normalized == normalized_area:
        return True
    return False


def ensure_master_budget_areas() -> dict[str, AreaPresupuesto]:
    areas: dict[str, AreaPresupuesto] = {}
    for code, name, order in AREA_DEFINITIONS:
        area, _ = AreaPresupuesto.objects.update_or_create(
            codigo=code,
            defaults={"nombre": name, "orden": order, "activa": True},
        )
        areas[code] = area
    return areas


def month_periods(year: int) -> list[date]:
    return [date(year, month_number, 1) for _, month_number in MONTH_COLUMNS]


class PresupuestoMaestroImportService:
    def _normalize_cell(self, value: object) -> str:
        return normalize_concept_text(value)

    def _csv_rows(self, path: Path) -> Iterable[dict[str, object]]:
        with path.open("r", newline="", encoding="utf-8-sig") as fh:
            yield from csv.DictReader(fh)

    def _xlsx_rows(self, path: Path) -> Iterable[dict[str, object]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        general_sheet = self._find_general_sales_sheet(workbook)
        if general_sheet is not None:
            yield from self._sales_projection_rows(general_sheet)
            return
        for sheet in workbook.worksheets:
            yield from self._sales_projection_rows(sheet)
            yield from self._budget_table_rows(sheet)

    def _find_general_sales_sheet(self, workbook):
        for sheet in workbook.worksheets:
            if self._normalize_cell(sheet.title) == "general" and self._sales_projection_layout(sheet):
                return sheet
        return None

    def _sales_projection_rows(self, sheet) -> Iterable[dict[str, object]]:
        rows = list(sheet.iter_rows(values_only=True))
        layout = self._sales_projection_layout(sheet, rows=rows)
        if layout is None:
            return

        branch_name = ""
        current_parent = ""
        for row in rows[layout["data_start"] :]:
            concept = self._sales_concept_from_row(row)
            if is_totalizer_budget_concept(concept, area_code="ventas"):
                continue
            if normalize_header_text(concept) in SALES_PARENT_CONCEPTS:
                current_parent = concept
                continue
            output_concept = f"{current_parent} · {concept}" if current_parent else concept
            output = {
                "concepto": output_concept,
                "tipo": RubroPresupuesto.TIPO_INGRESO,
                "sucursal": branch_name,
                "codigo_cuenta": "",
            }
            has_value = False
            for month_name, col_idx in layout["budget_columns"].items():
                value = row[col_idx] if col_idx < len(row) else None
                output[month_name] = value
                has_value = has_value or money(value) != 0
            for month_name, col_idx in layout["actual_columns"].items():
                value = row[col_idx] if col_idx < len(row) else None
                output[f"{month_name}_real"] = value
            for month_name, col_idx in layout.get("qty_columns", {}).items():
                value = row[col_idx] if col_idx < len(row) else None
                output[f"{month_name}_qty"] = value
            if has_value:
                yield output

    def _sales_projection_layout(self, sheet, *, rows: list[tuple[object, ...]] | None = None) -> dict[str, object] | None:
        rows = rows if rows is not None else list(sheet.iter_rows(values_only=True))
        if len(rows) < 4:
            return None

        for month_row_idx in range(min(12, len(rows))):
            month_row = rows[month_row_idx]
            month_starts: list[tuple[str, int]] = []
            boundaries: list[int] = []
            for idx, value in enumerate(month_row):
                month_name = normalize_header_text(value)
                if month_name in MONTH_NAME_TO_NUMBER:
                    month_starts.append((month_name, idx))
                    boundaries.append(idx)
                elif month_name in {"anual", "costo"} or month_name.startswith("%"):
                    boundaries.append(idx)
            if len(month_starts) < 3:
                continue
            boundaries = sorted(set(boundaries + [len(month_row)]))

            header_row_indexes = list(range(month_row_idx + 1, min(month_row_idx + 5, len(rows))))
            budget_columns: dict[str, int] = {}
            actual_columns: dict[str, int] = {}
            qty_columns: dict[str, int] = {}
            for pos, (month_name, start_idx) in enumerate(month_starts):
                end_idx = next((boundary for boundary in boundaries if boundary > start_idx), len(month_row))
                for col_idx in range(start_idx, end_idx):
                    if self._is_sales_value_column(rows, header_row_indexes, col_idx, kind="budget"):
                        budget_columns[month_name] = col_idx
                    if self._is_sales_value_column(rows, header_row_indexes, col_idx, kind="actual"):
                        actual_columns[month_name] = col_idx
                    if self._is_sales_value_column(rows, header_row_indexes, col_idx, kind="budget_qty"):
                        qty_columns[month_name] = col_idx
                if month_name not in budget_columns:
                    continue

            if budget_columns:
                last_header_idx = month_row_idx
                for row_idx in header_row_indexes:
                    labels = [normalize_header_text(value) for value in rows[row_idx]]
                    if any(
                        label in {"cant", "cantidad", "venta", "dif"}
                        or "proy" in label
                        or "proyeccion" in label
                        or "result" in label
                        or "2026" in label
                        for label in labels
                    ):
                        last_header_idx = row_idx
                return {
                    "data_start": min(last_header_idx + 1, len(rows)),
                    "budget_columns": budget_columns,
                    "actual_columns": actual_columns,
                    "qty_columns": qty_columns,
                }
        return None

    def _is_sales_value_column(
        self,
        rows: list[tuple[object, ...]],
        header_row_indexes: list[int],
        col_idx: int,
        *,
        kind: str,
    ) -> bool:
        current_labels = []
        previous_labels = []
        for row_idx in header_row_indexes:
            row = rows[row_idx]
            current_labels.append(normalize_header_text(row[col_idx] if col_idx < len(row) else ""))
            previous_labels.append(normalize_header_text(row[col_idx - 1] if col_idx > 0 and col_idx - 1 < len(row) else ""))
        labels = " ".join(label for label in current_labels + previous_labels if label)
        current = " ".join(label for label in current_labels if label)
        if kind == "budget_qty":
            # Columna CANTIDAD de la proyección (unidades propuestas del mes).
            is_qty = "cant" in current
            return is_qty and ("proy" in labels or "proyeccion" in labels)
        is_sale_amount = "venta" in current and "cant" not in current and "cantidad" not in current
        if not is_sale_amount:
            return False
        if kind == "budget":
            return "proy" in labels or "proyeccion" in labels
        return ("result" in labels or "real" in labels) and "2026" in labels

    def _sales_concept_from_row(self, row: tuple[object, ...]) -> str:
        for value in list(row[:4]):
            concept = str(value or "").strip()
            if not concept:
                continue
            normalized = normalize_header_text(concept)
            if normalized.replace(".", "", 1).isdigit():
                continue
            if normalized in {"cant", "cantidad", "venta", "dif", "diferencia"}:
                continue
            return concept
        return str(row[0] or "").strip() if row else ""

    def _budget_table_rows(self, sheet) -> Iterable[dict[str, object]]:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return

        for header_idx in range(min(8, len(rows))):
            budget_columns = self._month_budget_columns(rows, header_idx)
            if not budget_columns:
                continue
            concept_col = self._concept_column(rows, header_idx)
            if concept_col is None:
                continue
            account_col = self._account_column(rows, header_idx)
            branch_name = self._branch_from_sheet(sheet.title)
            for row in rows[header_idx + 1 :]:
                concept = str(row[concept_col] or "").strip() if concept_col < len(row) else ""
                if is_totalizer_budget_concept(concept, area_code=sheet.title):
                    continue
                output = {
                    "concepto": concept,
                    "tipo": "",
                    "sucursal": branch_name,
                    "codigo_cuenta": str(row[account_col] or "").strip() if account_col is not None and account_col < len(row) else "",
                }
                has_value = False
                for month_name, col_idx in budget_columns.items():
                    value = row[col_idx] if col_idx < len(row) else None
                    output[month_name] = value
                    has_value = has_value or money(value) != 0
                if has_value:
                    yield output
            return

    def _month_budget_columns(self, rows: list[tuple[object, ...]], header_idx: int) -> dict[str, int]:
        header = rows[header_idx]
        next_row = rows[header_idx + 1] if header_idx + 1 < len(rows) else ()
        columns: dict[str, int] = {}

        for idx, value in enumerate(header):
            label = self._normalize_cell(value)
            for month_name in MONTH_NAME_TO_NUMBER:
                if label == month_name or label == f"{month_name} presupuestado":
                    if "presupuestado" in label:
                        columns[month_name] = idx
                    elif idx < len(next_row) and self._normalize_cell(next_row[idx]) == "presupuestado":
                        columns[month_name] = idx

        if columns:
            return columns

        for idx, value in enumerate(header):
            label = self._normalize_cell(value)
            if label in MONTH_NAME_TO_NUMBER:
                columns[label] = idx
        return columns if len(columns) >= 3 else {}

    def _concept_column(self, rows: list[tuple[object, ...]], header_idx: int) -> int | None:
        candidates = [rows[header_idx]]
        if header_idx + 1 < len(rows):
            candidates.append(rows[header_idx + 1])
        accepted = {"concepto", "conceptos", "descripcion", "descripción", "producto/insumo", "unidad de negocio"}
        for row in candidates:
            for idx, value in enumerate(row):
                if self._normalize_cell(value) in accepted:
                    return idx
        return 1 if len(rows[header_idx]) > 1 else 0

    def _account_column(self, rows: list[tuple[object, ...]], header_idx: int) -> int | None:
        candidates = [rows[header_idx]]
        if header_idx + 1 < len(rows):
            candidates.append(rows[header_idx + 1])
        for row in candidates:
            for idx, value in enumerate(row):
                if self._normalize_cell(value) in {"cuenta", "clave"}:
                    return idx
        return None

    def _branch_from_sheet(self, title: str) -> str:
        normalized = self._normalize_cell(title)
        if normalized in {
            "general",
            "admon",
            "administracion",
            "produccion",
            "producccion",
            "logistica",
            "logística",
            "mantenimiento",
            "costo de produccion",
            "costo de producción",
            "presupuesto produccion",
            "presupuesto producción",
        }:
            return ""
        return title

    def _rows(self, path: Path) -> Iterable[dict[str, object]]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self._csv_rows(path)
        if suffix in {".xlsx", ".xlsm"}:
            return self._xlsx_rows(path)
        raise ValueError("Formato no soportado. Usa CSV o XLSX.")

    def _resolve_branch(self, value: object) -> Sucursal | None:
        # FASE 2: resolver canónico único (por nombre/código normalizado, tolerante a
        # prefijo 'Sucursal ' y acentos), en vez de igualdad/contains de texto frágil.
        # Índice cacheado por instancia para no re-consultar el catálogo por fila.
        raw = str(value or "").strip()
        if not raw:
            return None
        indice = getattr(self, "_branch_index", None)
        if indice is None:
            indice = self._branch_index = indice_sucursales_por_texto()
        return resolver_sucursal_por_texto(raw, indice=indice)

    @transaction.atomic
    def import_file(
        self,
        *,
        archivo: str | Path,
        area_code: str,
        version: str = LineaPresupuestoMensual.VERSION_ORIGINAL,
        year: int = 2026,
        source_name: str = "",
    ) -> PresupuestoImportSummary:
        path = Path(archivo).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(path)

        areas = ensure_master_budget_areas()
        normalized_area = normalize_area_code(area_code)
        if normalized_area not in areas:
            raise ValueError(f"Área de presupuesto desconocida: {area_code}")
        area = areas[normalized_area]
        version = normalize_version(version)

        rubros_created = 0
        rubros_updated = 0
        lines_created = 0
        lines_updated = 0
        skipped_rows = 0
        source = source_name or path.stem.upper()

        for row in self._rows(path):
            concept = str(row.get("concepto") or row.get("concept") or "").strip()
            if is_totalizer_budget_concept(concept, area_code=normalized_area):
                skipped_rows += 1
                continue
            branch = self._resolve_branch(row.get("sucursal"))
            account_code = str(row.get("codigo_cuenta") or row.get("cuenta") or "").strip()
            rubro_type = infer_rubro_type(row.get("tipo"), normalized_area, concept)
            rubro, was_created = RubroPresupuesto.objects.update_or_create(
                area=area,
                concepto=concept,
                codigo_cuenta=account_code,
                sucursal=branch,
                defaults={
                    "tipo": rubro_type,
                    "activo": True,
                    "metadata": {
                        "source": source,
                        "source_file": path.name,
                        "actual_key": normalize_actual_key(concept, account_code, area.codigo),
                    },
                },
            )
            rubros_created += int(was_created)
            rubros_updated += int(not was_created)
            for month_name, month_number in MONTH_COLUMNS:
                period = date(year, month_number, 1)
                amount = money(row.get(month_name))
                existente = LineaPresupuestoMensual.objects.filter(
                    rubro=rubro, periodo=period, version=version
                ).first()

                # El importador de presupuesto SOLO manda sobre monto_presupuesto.
                # La metadata se fusiona (no se reemplaza) para no destruir el
                # historial de capturas ni el desglose de la consolidación, y el
                # real del archivo (columna _real) jamás pisa una captura
                # MANUAL:* ni un consolidado AUTO:* vigente.
                metadata = dict(existente.metadata or {}) if existente else {}
                metadata["source"] = source
                metadata["source_file"] = path.name
                line_defaults = {"monto_presupuesto": amount, "metadata": metadata}

                if f"{month_name}_real" in row:
                    fuente_actual = str(existente.fuente_real or "") if existente else ""
                    protegida = fuente_actual.startswith("MANUAL:") or fuente_actual.startswith("AUTO:")
                    if not protegida:
                        line_defaults.update(
                            {
                                # Namespace re-escribible: la consolidación viva
                                # (POS/nómina) puede reemplazarlo después.
                                "monto_real": money(row.get(f"{month_name}_real")),
                                "fuente_real": "AUTO:LEGADO",
                            }
                        )
                        metadata["fuente_import"] = source

                _, line_created = LineaPresupuestoMensual.objects.update_or_create(
                    rubro=rubro,
                    periodo=period,
                    version=version,
                    defaults=line_defaults,
                )
                lines_created += int(line_created)
                lines_updated += int(not line_created)

        return PresupuestoImportSummary(
            area=area.codigo,
            version=version,
            year=year,
            rubros_created=rubros_created,
            rubros_updated=rubros_updated,
            lines_created=lines_created,
            lines_updated=lines_updated,
            skipped_rows=skipped_rows,
        )

    def reimport_sales_projection(
        self,
        *,
        archivo: str | Path,
        version: str = LineaPresupuestoMensual.VERSION_ORIGINAL,
        year: int = 2026,
        source_name: str = "",
        clear_first: bool = False,
        dry_run: bool = False,
    ) -> VentasReimportSummary:
        path = Path(archivo).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(path)

        version = normalize_version(version)

        if dry_run:
            parsed_rows = list(self._rows(path))
            unique_concepts = {
                (
                    str(row.get("concepto") or "").strip(),
                    str(row.get("codigo_cuenta") or row.get("cuenta") or "").strip(),
                    str(row.get("sucursal") or "").strip(),
                )
                for row in parsed_rows
            }
            monthly_totals = {
                month_name: sum((money(row.get(month_name)) for row in parsed_rows), Decimal("0"))
                for month_name, _month_number in MONTH_COLUMNS
            }
            return VentasReimportSummary(
                year=year,
                version=version,
                dry_run=True,
                deleted_rubros=0,
                deleted_lines=0,
                rubros_created=len(unique_concepts),
                rubros_updated=0,
                lines_created=len(unique_concepts) * 12,
                lines_updated=0,
                skipped_rows=0,
                monthly_totals=monthly_totals,
            )

        areas = ensure_master_budget_areas()
        area = areas["ventas"]
        deleted_rubros = 0
        deleted_lines = 0

        with transaction.atomic():
            if clear_first:
                from .models import ReglaFuenteRubro

                rubros_qs = RubroPresupuesto.objects.filter(area=area)
                protegidas = (
                    LineaPresupuestoMensual.objects.filter(rubro__in=rubros_qs)
                    .filter(
                        Q(fuente_real__startswith="MANUAL:")
                        | Q(fuente_real__startswith="AUTO:")
                    )
                    .count()
                )
                reglas_admin = ReglaFuenteRubro.objects.filter(
                    rubro__in=rubros_qs, origen=ReglaFuenteRubro.ORIGEN_ADMIN
                ).count()
                if protegidas or reglas_admin:
                    raise ValueError(
                        f"El área tiene {protegidas} línea(s) con real protegido "
                        f"(MANUAL/AUTO) y {reglas_admin} regla(s) de mapeo ADMIN; "
                        "clear_first las destruiría. Resuélvelas antes de limpiar."
                    )
                deleted_rubros = rubros_qs.count()
                deleted_lines = LineaPresupuestoMensual.objects.filter(rubro__in=rubros_qs).count()
                LineaPresupuestoMensual.objects.filter(rubro__in=rubros_qs).delete()
                rubros_qs.delete()

            summary = self.import_file(
                archivo=path,
                area_code="ventas",
                version=version,
                year=year,
                source_name=source_name or path.stem.upper(),
            )
            unidades_upsertadas, unidades_sin_receta = self._upsert_pronostico_unidades(
                path=path, year=year
            )
            monthly_totals = {
                month_name: money(
                    LineaPresupuestoMensual.objects.filter(
                        rubro__area=area,
                        rubro__tipo=RubroPresupuesto.TIPO_INGRESO,
                        periodo=date(year, month_number, 1),
                        version=version,
                    ).aggregate(total=Sum("monto_presupuesto"))["total"]
                )
                for month_name, month_number in MONTH_COLUMNS
            }

        return VentasReimportSummary(
            year=year,
            version=version,
            dry_run=dry_run,
            deleted_rubros=deleted_rubros,
            deleted_lines=deleted_lines,
            rubros_created=summary.rubros_created,
            rubros_updated=summary.rubros_updated,
            lines_created=summary.lines_created,
            lines_updated=summary.lines_updated,
            skipped_rows=summary.skipped_rows,
            monthly_totals=monthly_totals,
            unidades_upsertadas=unidades_upsertadas,
            unidades_sin_receta=tuple(unidades_sin_receta),
        )

    def _upsert_pronostico_unidades(self, *, path: Path, year: int) -> tuple[int, list[str]]:
        """Guarda las CANTIDADES proyectadas del Excel en recetas.PronosticoVenta.

        El concepto del Excel se cruza con la Receta por matching difuso (el
        mismo criterio que el seed de reglas POS). Solo se pisan pronósticos
        cuya fuente sea de presupuesto (PRESUPUESTO_*) o esté vacía — un
        pronóstico capturado manualmente en recetas no se toca.
        """
        from rapidfuzz import fuzz

        from recetas.models import PronosticoVenta, Receta
        from reportes.management.commands.seed_reglas_fuente_rubro import canon_pos

        # Receta no tiene campo "activo"; se cruzan todas (fabricado y reventa).
        recetas = [(receta, canon_pos(receta.nombre)) for receta in Receta.objects.all()]

        # Cantidades por concepto×mes (la hoja GENERAL ya es toda la empresa).
        cantidades: dict[str, dict[int, Decimal]] = {}
        for row in self._rows(path):
            concepto = str(row.get("concepto") or "").strip()
            if not concepto:
                continue
            for month_name, month_number in MONTH_COLUMNS:
                qty = row.get(f"{month_name}_qty")
                if qty in (None, ""):
                    continue
                try:
                    valor = Decimal(str(qty))
                except Exception:  # noqa: BLE001 — celdas con texto se ignoran
                    continue
                if valor == 0:
                    continue
                cantidades.setdefault(concepto, {})
                cantidades[concepto][month_number] = (
                    cantidades[concepto].get(month_number, Decimal("0")) + valor
                )

        fuente = f"PRESUPUESTO_{year}"
        upsertadas = 0
        sin_receta: list[str] = []
        for concepto, meses in cantidades.items():
            objetivo = canon_pos(concepto)
            candidatos = [(receta, fuzz.token_set_ratio(objetivo, canon)) for receta, canon in recetas]
            mejor = max(candidatos, key=lambda c: c[1], default=None)
            if mejor is None or mejor[1] < 90:
                sin_receta.append(concepto)
                continue
            receta = mejor[0]
            for month_number, cantidad in meses.items():
                periodo = f"{year}-{month_number:02d}"
                existente = PronosticoVenta.objects.filter(receta=receta, periodo=periodo).first()
                if existente is not None and not (
                    not existente.fuente or existente.fuente.startswith("PRESUPUESTO")
                ):
                    continue  # capturado manualmente en recetas: no pisar
                PronosticoVenta.objects.update_or_create(
                    receta=receta,
                    periodo=periodo,
                    defaults={"cantidad": cantidad, "fuente": fuente},
                )
                upsertadas += 1
        return upsertadas, sin_receta


class PresupuestoMaestroService:
    AREA_ACTUAL_KEYS = {
        "ventas": "ventas",
        "produccion": "costo_mp",
        "compras": "costo_reventa",
        "administracion": "gasto_fijo",
        "nomina": "mano_obra",
    }

    def actuals_for_period(self, period: date) -> dict[str, Decimal]:
        result = EmpresaResultadoMensual.objects.filter(periodo=period).first()
        if result is None:
            return {}
        return {
            "ventas": money(result.venta_total),
            "costo_mp": money(result.costo_materia_prima_total),
            "costo_reventa": money(result.costo_reventa_total),
            "gasto_fijo": money(result.gasto_comercial_total) + money(result.gasto_corporativo_total),
            "mano_obra": money(result.mano_obra_prod_total),
            "indirectos": money(result.indirecto_prod_total) + money(result.empaque_prod_total),
            "utilidad_operativa": money(result.utilidad_operativa_total),
        }

    def _signed_budget(self, amount: Decimal, rubro_type: str) -> Decimal:
        value = money(amount)
        if rubro_type in {
            RubroPresupuesto.TIPO_EGRESO,
            RubroPresupuesto.TIPO_COSTO,
            RubroPresupuesto.TIPO_CAPEX,
        }:
            return -value
        return value

    def _budget_lines_for_kpi(self, *, year: int, version: str, area: str | None = None):
        area_code = normalize_area_code(area or "")
        qs = LineaPresupuestoMensual.objects.filter(periodo__year=year, version=version).select_related(
            "rubro", "rubro__area"
        )
        if area_code:
            return qs.filter(rubro__area__codigo=area_code, rubro__activo=True)
        return qs.filter(
            rubro__area__codigo="ventas",
            rubro__tipo=RubroPresupuesto.TIPO_INGRESO,
            rubro__activo=True,
        )

    def executive_kpis(
        self,
        *,
        year: int,
        month: int,
        version: str = LineaPresupuestoMensual.VERSION_ORIGINAL,
        area: str | None = None,
    ) -> dict[str, object]:
        version = normalize_version(version)
        area_code = normalize_area_code(area or "")
        month = max(1, min(int(month or 1), 12))
        period = date(year, month, 1)
        annual_budget = Decimal("0")
        monthly_budget = Decimal("0")
        for line in self._budget_lines_for_kpi(year=year, version=version, area=area_code or None):
            amount = money(line.monto_presupuesto)
            signed_amount = self._signed_budget(amount, line.rubro.tipo) if area_code else amount
            annual_budget += signed_amount
            if line.periodo == period:
                monthly_budget += signed_amount

        # El real del KPI debe ser la métrica DE ESA ÁREA, no siempre ventas
        # (hallazgo de auditoría: Nómina/Producción se comparaban contra
        # venta_total). Área sin métrica mapeada = sin dato, no un dato ajeno.
        actuals = self.actuals_for_period(period)
        actual_key = self.AREA_ACTUAL_KEYS.get(area_code, "ventas") if area_code else "ventas"
        real_month = actuals.get(actual_key)
        sin_dato = real_month is None
        if area_code and area_code not in self.AREA_ACTUAL_KEYS:
            real_month = None
            sin_dato = True
        if area_code:
            # KPI de área usa presupuesto con signo; el real de un área de
            # gasto también se compara en su propia escala (positivo).
            real_month = real_month if real_month is not None else Decimal("0")
            if actual_key != "ventas":
                real_month = -real_month
        else:
            real_month = real_month if real_month is not None else Decimal("0")
        variance = real_month - monthly_budget
        return {
            "year": year,
            "month": month,
            "periodo": period,
            "area": area_code,
            "annual_budget": annual_budget,
            "monthly_budget": monthly_budget,
            "real_month": real_month,
            "variance": variance,
            "variance_pct": variance_pct(variance, monthly_budget),
            "real_source": "" if sin_dato else "reportes.EmpresaResultadoMensual",
            "real_note": "Sin dato para esta área" if sin_dato else "",
            "budget_scope": "area_signed" if area_code else "ventas_ingreso",
        }

    def _line_actual(self, line: LineaPresupuestoMensual, actuals: dict[str, Decimal], consumed: set[str]) -> tuple[Decimal | None, str]:
        rubro = line.rubro
        # El real consolidado (AUTO:*) o capturado (MANUAL:*) tiene precedencia
        # sobre el fallback heurístico de EmpresaResultadoMensual.
        fuente = str(line.fuente_real or "")
        if line.monto_real is not None and (fuente.startswith("AUTO:") or fuente.startswith("MANUAL:")):
            return line.monto_real, line.fuente_real
        actual_key = str((rubro.metadata or {}).get("actual_key") or "")
        if not actual_key:
            actual_key = normalize_actual_key(rubro.codigo_cuenta, rubro.concepto, rubro.area.codigo)
        if rubro.sucursal_id:
            return None, ""
        if actual_key and actual_key in actuals and actual_key not in consumed:
            consumed.add(actual_key)
            return actuals[actual_key], "reportes.EmpresaResultadoMensual"
        return line.monto_real, line.fuente_real

    def build_consolidado(
        self,
        *,
        periodo: str | date,
        version: str = LineaPresupuestoMensual.VERSION_ORIGINAL,
        area: str | None = None,
    ) -> dict[str, object]:
        period = parse_period(periodo)
        version = normalize_version(version)
        ensure_master_budget_areas()
        actuals = self.actuals_for_period(period)
        areas_qs = AreaPresupuesto.objects.filter(activa=True).order_by("orden", "nombre")
        if area:
            areas_qs = areas_qs.filter(codigo=normalize_area_code(area))
        lines = (
            LineaPresupuestoMensual.objects.filter(periodo=period, version=version, rubro__area__in=areas_qs)
            .select_related("rubro", "rubro__area", "rubro__sucursal")
            .order_by("rubro__area__orden", "rubro__concepto", "rubro__sucursal__codigo")
        )

        consumed_actuals: set[str] = set()
        area_map: dict[int, dict[str, object]] = {}
        total_budget = Decimal("0")
        total_actual = Decimal("0")
        total_variance = Decimal("0")

        for line in lines:
            rubro = line.rubro
            area_payload = area_map.setdefault(
                rubro.area_id,
                {
                    "id": rubro.area_id,
                    "codigo": rubro.area.codigo,
                    "nombre": rubro.area.nombre,
                    "orden": rubro.area.orden,
                    "rubros": [],
                    "total_presupuesto": Decimal("0"),
                    "total_presupuesto_signed": Decimal("0"),
                    "total_real": Decimal("0"),
                    "total_varianza": Decimal("0"),
                },
            )
            actual, fuente_real = self._line_actual(line, actuals, consumed_actuals)
            budget = money(line.monto_presupuesto)
            signed_budget = self._signed_budget(budget, rubro.tipo)
            actual_value = money(actual) if actual is not None else Decimal("0")
            variance = actual_value - budget
            pct = variance_pct(variance, budget)
            tone = self._tone(rubro.tipo, variance, pct)
            row = {
                "linea_id": line.id,
                "rubro_id": rubro.id,
                "concepto": rubro.concepto,
                "codigo_cuenta": rubro.codigo_cuenta,
                "tipo": rubro.tipo,
                "tipo_label": rubro.get_tipo_display(),
                "type_class": self._type_class(rubro.tipo),
                "sucursal": rubro.sucursal.codigo if rubro.sucursal_id else "",
                "presupuesto": budget,
                "presupuesto_signed": signed_budget,
                "real": actual_value if actual is not None else None,
                "fuente_real": fuente_real,
                "varianza": variance if actual is not None else None,
                "varianza_pct": pct if actual is not None else None,
                "tone": tone,
            }
            area_payload["rubros"].append(row)
            area_payload["total_presupuesto"] += budget
            area_payload["total_presupuesto_signed"] += signed_budget
            total_budget += budget
            if actual is not None:
                area_payload["total_real"] += actual_value
                area_payload["total_varianza"] += variance
                total_actual += actual_value
                total_variance += variance

        areas = list(area_map.values())
        for area_payload in areas:
            actual_key = self.AREA_ACTUAL_KEYS.get(str(area_payload["codigo"]))
            if actual_key and actual_key in actuals:
                previous_actual = money(area_payload["total_real"])
                canonical_actual = money(actuals[actual_key])
                previous_variance = money(area_payload["total_varianza"])
                canonical_variance = canonical_actual - money(area_payload["total_presupuesto"])
                area_payload["total_real"] = canonical_actual
                area_payload["total_varianza"] = canonical_variance
                total_actual += canonical_actual - previous_actual
                total_variance += canonical_variance - previous_variance
                area_payload["fuente_real"] = "reportes.EmpresaResultadoMensual"
            area_payload["varianza_pct"] = variance_pct(area_payload["total_varianza"], area_payload["total_presupuesto"])
            area_payload["tone"] = "neutral"
            area_payload["tipo_mix"] = self._area_type_mix(area_payload["rubros"])

        return {
            "periodo": period,
            "version": version,
            "areas": areas,
            "totales": {
                "presupuesto": total_budget,
                "real": total_actual,
                "varianza": total_variance,
                "varianza_pct": variance_pct(total_variance, total_budget),
            },
            "actual_source": "reportes.EmpresaResultadoMensual",
            "actuals": {key: value for key, value in actuals.items()},
            "actual_labels": ACTUAL_LABELS,
        }

    def _area_type_mix(self, rows: list[dict[str, object]]) -> str:
        types = {str(row.get("tipo") or "") for row in rows}
        if len(types) == 1:
            return next(iter(types), "")
        return "MIXTO"

    def annual_matrix(self, *, year: int, version: str, area: str | None = None) -> dict[str, object]:
        version = normalize_version(version)
        ensure_master_budget_areas()
        areas_qs = AreaPresupuesto.objects.filter(activa=True).order_by("orden", "nombre")
        if area:
            areas_qs = areas_qs.filter(codigo=normalize_area_code(area))
        rubros = RubroPresupuesto.objects.filter(area__in=areas_qs, activo=True).select_related("area", "sucursal").order_by(
            "area__orden", "area__nombre", "concepto", "sucursal__codigo"
        )
        if normalize_area_code(area or "") == "ventas":
            rubros = rubros.filter(tipo=RubroPresupuesto.TIPO_INGRESO)
        lines_by_rubro_period = {
            (line.rubro_id, line.periodo.month): line
            for line in LineaPresupuestoMensual.objects.filter(periodo__year=year, version=version, rubro__in=rubros)
        }
        rows = []
        area_totals: dict[int, dict[str, object]] = {}
        total_anual = Decimal("0")
        for rubro in rubros:
            month_cells = []
            total = Decimal("0")
            signed_total = Decimal("0")
            for month_name, month_number in MONTH_COLUMNS:
                line = lines_by_rubro_period.get((rubro.id, month_number))
                amount = money(line.monto_presupuesto if line else Decimal("0"))
                total += amount
                signed_total += self._signed_budget(amount, rubro.tipo)
                month_cells.append(
                    {
                        "name": month_name,
                        "number": month_number,
                        "linea_id": line.id if line else None,
                        "amount": amount,
                    }
                )
            if total == 0:
                continue
            total_anual += total
            area_total = area_totals.setdefault(
                rubro.area_id,
                {
                    "area": rubro.area,
                    "total": Decimal("0"),
                    "signed_total": Decimal("0"),
                    "row_count": 0,
                },
            )
            area_total["total"] += total
            area_total["signed_total"] += signed_total
            area_total["row_count"] += 1
            rows.append(
                {
                    "rubro": rubro,
                    "area": rubro.area,
                    "months": month_cells,
                    "total": total,
                    "signed_total": signed_total,
                    "type_class": self._type_class(rubro.tipo),
                }
            )
        return {
            "year": year,
            "version": version,
            "area": area or "",
            "months": MONTH_COLUMNS,
            "areas": list(areas_qs),
            "rows": rows,
            "area_totals": sorted(area_totals.values(), key=lambda row: (row["area"].orden, row["area"].nombre)),
            "total_anual": total_anual,
        }

    def _type_class(self, rubro_type: str) -> str:
        return {
            RubroPresupuesto.TIPO_INGRESO: "is-income",
            RubroPresupuesto.TIPO_EGRESO: "is-expense",
            RubroPresupuesto.TIPO_COSTO: "is-cost",
            RubroPresupuesto.TIPO_CAPEX: "is-capex",
        }.get(rubro_type, "is-mixed")

    def update_line_amount(self, *, line_id: int, amount: Decimal) -> LineaPresupuestoMensual:
        line = LineaPresupuestoMensual.objects.select_related("rubro").get(pk=line_id)
        line.monto_presupuesto = money(amount)
        line.save(update_fields=["monto_presupuesto", "actualizado_en"])
        return line

    @transaction.atomic
    def create_rubro_with_empty_year(
        self,
        *,
        area_code: str,
        concepto: str,
        tipo: str,
        year: int,
        version: str,
        codigo_cuenta: str = "",
        sucursal_id: int | None = None,
    ) -> RubroPresupuesto:
        areas = ensure_master_budget_areas()
        area = areas[normalize_area_code(area_code)]
        branch = Sucursal.objects.filter(pk=sucursal_id).first() if sucursal_id else None
        rubro, _ = RubroPresupuesto.objects.update_or_create(
            area=area,
            concepto=concepto.strip(),
            codigo_cuenta=codigo_cuenta.strip(),
            sucursal=branch,
            defaults={
                "tipo": normalize_rubro_type(tipo, area.codigo),
                "activo": True,
                "metadata": {"source": "ui", "actual_key": normalize_actual_key(concepto, codigo_cuenta, area.codigo)},
            },
        )
        for period in month_periods(year):
            LineaPresupuestoMensual.objects.update_or_create(
                rubro=rubro,
                periodo=period,
                version=normalize_version(version),
                defaults={"monto_presupuesto": Decimal("0"), "metadata": {"source": "ui"}},
            )
        return rubro

    def _tone(self, rubro_type: str, variance: Decimal, pct: Decimal) -> str:
        abs_pct = abs(pct)
        if variance == 0 or pct == 0:
            return "neutral"
        favorable = variance > 0 if rubro_type == RubroPresupuesto.TIPO_INGRESO else variance < 0
        if favorable:
            return "success"
        if abs_pct > Decimal("10"):
            return "danger"
        if abs_pct >= Decimal("5"):
            return "warning"
        return "neutral"


def seed_capex_guamuchil_2026() -> dict[str, int]:
    areas = ensure_master_budget_areas()
    area = areas["capex"]
    created = 0
    updated = 0
    for concept, year, month, amount in CAPEX_GUAMUCHIL_ROWS:
        rubro, rubro_created = RubroPresupuesto.objects.update_or_create(
            area=area,
            concepto=concept,
            codigo_cuenta="CAPEX_GUAMUCHIL",
            sucursal=None,
            defaults={
                "tipo": RubroPresupuesto.TIPO_CAPEX,
                "activo": True,
                "metadata": {"source": "seed_presupuesto_2026", "confirmed_real": True},
            },
        )
        _, line_created = LineaPresupuestoMensual.objects.update_or_create(
            rubro=rubro,
            periodo=date(year, month, 1),
            version=LineaPresupuestoMensual.VERSION_ORIGINAL,
            defaults={
                "monto_presupuesto": amount,
                "monto_real": amount,
                "fuente_real": "CAPEX_GUAMUCHIL_CONFIRMADO",
                "metadata": {"source": "seed_presupuesto_2026", "confirmed_real": True},
            },
        )
        created += int(rubro_created or line_created)
        updated += int(not line_created)
    return {
        "areas": AreaPresupuesto.objects.count(),
        "created": created,
        "updated": updated,
        "capex_total": int(
            LineaPresupuestoMensual.objects.filter(
                rubro__area=area,
                periodo__year=2026,
                version=LineaPresupuestoMensual.VERSION_ORIGINAL,
            ).aggregate(total=Sum("monto_presupuesto"))["total"]
            or 0
        ),
    }
