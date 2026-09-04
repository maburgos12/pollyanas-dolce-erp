from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
import re
from types import SimpleNamespace
from unittest.mock import call, patch

from django.test import RequestFactory, TestCase
from django.template.loader import render_to_string
from django.contrib.auth.models import AnonymousUser
from django.utils import timezone
from openpyxl import load_workbook

from pos_bridge.models import PointBranch, PointConversionLine, PointDailySale, PointProduct
from pos_bridge.services.monthly_product_balance_service import MonthlyPointBalanceRow
from recetas.models import Receta
from reportes.views_produccion import ProducidoVsVendidoMermaView


ZERO = Decimal("0")


def canonical_balance(*rows, sources=None, warnings=(), issues=()):
    return SimpleNamespace(
        rows={row.receta_id: row for row in rows},
        sources=sources
        or {
            "opening_snapshot": {
                "source": "PointInventorySnapshot",
                "authoritative": True,
                "effective_date": date(2026, 7, 31),
                "applied_coverage_key_count": 9,
            },
            "closing_snapshot": {
                "source": "PointInventorySnapshot",
                "authoritative": True,
                "effective_date": date(2026, 8, 31),
                "applied_coverage_key_count": 9,
            },
            "production": {"source": "PointProductionLine", "authoritative": True, "source_present": True},
            "sales": {
                "source": "PointDailySale",
                "authoritative": True,
                "mode": "OFFICIAL_DAILY",
                "selected_source": "PointDailySale",
            },
            "waste": {"source": "PointWasteLine", "authoritative": True, "source_present": True},
            "conversions": {"source": "PointConversionLine", "authoritative": True, "source_present": True},
        },
        effective_snapshot_dates={"opening": date(2026, 7, 31), "closing": date(2026, 8, 31)},
        warnings=warnings,
        issues=issues,
        source_counts={"sales_rows": 1, "production_rows": 1},
    )


class ProducidoVsVendidoCanonicalBalanceTests(TestCase):
    def test_known_point_closing_is_displayed_even_when_opening_or_audit_is_missing(self):
        balance = canonical_balance(MonthlyPointBalanceRow(
            receta_id=self.parent.id, opening_point=None, closing_point=Decimal("22"),
            status="REVISAR_FUENTE",
        ))
        balance.sources["opening_snapshot"]["authoritative"] = False
        balance.sources["closing_snapshot"]["authoritative"] = False
        context, _ = self._context(balance)
        row = context["groups"][0]["rows"][0]
        self.assertEqual(row["inventario_final_point_total"], Decimal("22"))
        self.assertIsNone(row["inventario_inicial"])
        self.assertIsNone(row["diferencia_inventario"])
        self.assertEqual(row["estado_inventario"], "Revisar fuente")
        self.assertNotIn("los saldos y su diferencia se muestran como Sin dato", self._render(context))
        self.assertIn("fecha operativa", self._render(context))
        for export in ("_export_csv", "_export_xlsx", "_export_pdf"):
            response = getattr(ProducidoVsVendidoMermaView(), export)(context)
            self.assertEqual(response.status_code, 200)

    def setUp(self):
        self.factory = RequestFactory()
        self.parent = Receta.objects.create(
            nombre="Pastel canónico",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Pastel Mediano",
            hash_contenido="test-pvv-parent",
        )
        self.slice = Receta.objects.create(
            nombre="Rebanada canónica",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Rebanada",
            hash_contenido="test-pvv-slice",
        )

    def _context(self, balance, period="2026-08"):
        with patch("reportes.views_produccion.MonthlyPointProductBalanceService") as service:
            service.return_value.build.return_value = balance
            context = ProducidoVsVendidoMermaView()._build_context(
                self.factory.get("/reportes/produccion/", {"periodo": period})
            )
        return context, service

    def _render(self, context):
        request = self.factory.get("/reportes/produccion/")
        request.user = AnonymousUser()
        return render_to_string(
            "reportes/producido_vs_vendido.html",
            context,
            request=request,
        )

    def _conversion_origin_context(self):
        mixed = Receta.objects.create(
            nombre="Producto conversión mixta",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Galletas",
            hash_contenido="test-pvv-mixed",
        )
        unresolved = Receta.objects.create(
            nombre="Producto conversión sin resolver",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Pay Grande",
            hash_contenido="test-pvv-unresolved",
        )
        context, _ = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(
                    receta_id=self.parent.id,
                    conversion_in=Decimal("1"),
                    conversion_origin="POINT",
                ),
                MonthlyPointBalanceRow(
                    receta_id=self.slice.id,
                    conversion_in=Decimal("1"),
                    conversion_origin="EQUIVALENCIA_CONFIGURADA",
                ),
                MonthlyPointBalanceRow(
                    receta_id=mixed.id,
                    conversion_in=Decimal("1"),
                    conversion_origin="MIXED",
                ),
                MonthlyPointBalanceRow(
                    receta_id=unresolved.id,
                    conversion_in=Decimal("1"),
                    conversion_origin="UNRESOLVED",
                ),
            )
        )
        return context

    def test_slice_sale_without_point_conversion_never_infers_conversion_from_sale_or_waste(self):
        context, _ = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(
                    receta_id=self.slice.id,
                    opening_point=Decimal("0"),
                    sales=Decimal("6"),
                    waste=Decimal("2"),
                    conversion_in=ZERO,
                    conversion_out=ZERO,
                    calculated_closing=Decimal("-8"),
                    closing_point=Decimal("-8"),
                    difference_point=ZERO,
                    status="COINCIDE",
                ),
                MonthlyPointBalanceRow(
                    receta_id=self.parent.id,
                    opening_point=Decimal("5"),
                    calculated_closing=Decimal("5"),
                    closing_point=Decimal("5"),
                    difference_point=ZERO,
                    status="COINCIDE",
                ),
            )
        )
        by_recipe = {row["receta_id"]: row for group in context["groups"] for row in group["rows"]}
        self.assertEqual(by_recipe[self.slice.id]["conversion_entrada"], ZERO)
        self.assertEqual(by_recipe[self.parent.id]["conversion_salida"], ZERO)

    def test_uses_canonical_movements_formula_difference_status_and_conversion_provenance(self):
        context, service = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(
                    receta_id=self.parent.id,
                    opening_point=Decimal("10"),
                    production=Decimal("5"),
                    sales=Decimal("4"),
                    waste=Decimal("1"),
                    conversion_in=Decimal("2"),
                    conversion_out=Decimal("3"),
                    calculated_closing=Decimal("9"),
                    closing_point=Decimal("11"),
                    difference_point=Decimal("2"),
                    status="POINT_MAYOR",
                    conversion_origin="point_conversion_line",
                )
            )
        )
        row = context["groups"][0]["rows"][0]
        self.assertEqual(row["inventario_inicial"], Decimal("10"))
        self.assertEqual(row["inventario_final_teorico"], Decimal("9"))
        self.assertEqual(row["inventario_final_point_total"], Decimal("11"))
        self.assertEqual(row["diferencia_inventario"], Decimal("2"))
        self.assertEqual(row["estado_inventario"], "Point mayor")
        self.assertEqual(row["conversion_entrada"], Decimal("2"))
        self.assertEqual(row["conversion_salida"], Decimal("3"))
        self.assertEqual(row["conversion_provenance"], "point_conversion_line")
        service.return_value.build.assert_called_once_with(month="2026-08")

    def test_status_labels_are_neutral(self):
        self.assertEqual(ProducidoVsVendidoMermaView._point_status_label("COINCIDE"), "Coincide")
        self.assertEqual(ProducidoVsVendidoMermaView._point_status_label("POINT_MAYOR"), "Point mayor")
        self.assertEqual(ProducidoVsVendidoMermaView._point_status_label("POINT_MENOR"), "Point menor")
        self.assertEqual(ProducidoVsVendidoMermaView._point_status_label("REVISAR_FUENTE"), "Revisar fuente")

    def test_missing_or_non_authoritative_balance_is_unavailable_not_zero_or_match(self):
        context, _ = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(
                    receta_id=self.parent.id,
                    opening_point=None,
                    production=Decimal("5"),
                    sales=Decimal("4"),
                    calculated_closing=None,
                    closing_point=None,
                    difference_point=None,
                    status="REVISAR_FUENTE",
                ),
                sources={
                    "opening_snapshot": {"source": "PointInventorySnapshot", "authoritative": False},
                    "closing_snapshot": {"source": "PointInventorySnapshot", "authoritative": False},
                    "production": {"source": "PointProductionLine", "authoritative": True},
                    "sales": {"source": "PointDailySale", "authoritative": True},
                    "waste": {"source": "PointWasteLine", "authoritative": True},
                    "conversions": {"source": "PointConversionLine"},
                },
            )
        )
        row = context["groups"][0]["rows"][0]
        self.assertIsNone(row["inventario_inicial"])
        self.assertIsNone(row["inventario_final_teorico"])
        self.assertIsNone(row["inventario_final_point_total"])
        self.assertIsNone(row["diferencia_inventario"])
        self.assertEqual(row["estado_inventario"], "Revisar fuente")

    def test_totals_keep_authoritative_dif_when_reference_rows_share_the_category(self):
        reference = Receta.objects.create(
            nombre="Referencia canónica",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Pastel Mediano",
            pasa_modulo_produccion=False,
            hash_contenido="test-pvv-reference",
        )
        self.parent.pasa_modulo_produccion = True
        self.parent.save(update_fields=["pasa_modulo_produccion"])
        context, _ = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(
                    receta_id=self.parent.id,
                    production=Decimal("10"),
                    sales=Decimal("2"),
                    status="REVISAR_FUENTE",
                ),
                MonthlyPointBalanceRow(
                    receta_id=reference.id,
                    production=Decimal("20"),
                    sales=Decimal("1"),
                    status="REVISAR_FUENTE",
                ),
            )
        )
        self.assertEqual(context["groups"][0]["total"]["dif"], Decimal("8"))
        self.assertEqual(context["grand_total"]["dif"], Decimal("8"))

    def test_operational_difference_total_is_unavailable_when_any_operational_row_is_missing(self):
        total = ProducidoVsVendidoMermaView()._difference_total(
            [
                {"produccion_referencia": False, "dif": Decimal("2")},
                {"produccion_referencia": False, "dif": None},
                {"produccion_referencia": True, "dif": None},
            ]
        )

        self.assertIsNone(total)

    def test_missing_total_values_render_as_sin_dato_without_bare_currency_or_success_state(self):
        context, _ = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(receta_id=self.parent.id, status="REVISAR_FUENTE"),
                sources={
                    "opening_snapshot": {"source": "PointInventorySnapshot", "authoritative": False},
                    "closing_snapshot": {"source": "PointInventorySnapshot", "authoritative": False},
                    "production": {"source": "PointProductionLine", "source_present": False},
                    "sales": {"source": "PointDailySale", "source_present": False},
                    "waste": {"source": "PointWasteLine", "source_present": False},
                    "conversions": {"source": "PointConversionLine"},
                },
            )
        )
        self.assertIsNone(context["grand_total"]["vendido"])
        self.assertIsNone(context["grand_total"]["producido"])
        self.assertIsNone(context["grand_total"]["dif"])
        self.assertIsNone(context["grand_total"]["costo_merma"])
        rendered = self._render(context)
        self.assertGreaterEqual(rendered.count("Sin dato"), 8)
        self.assertNotIn('kpi-number is-success"></div>', rendered)
        self.assertNotIn('class="text-success"></span>', rendered)
        self.assertNotIn('>$</div>', rendered)

    def test_authoritative_zero_waste_has_zero_cost_even_without_known_unit_cost(self):
        context, _ = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(
                    receta_id=self.parent.id,
                    waste=ZERO,
                    status="COINCIDE",
                )
            )
        )

        row = context["groups"][0]["rows"][0]
        self.assertEqual(row["merma_reportada"], ZERO)
        self.assertEqual(row["costo_merma"], ZERO)
        self.assertEqual(context["grand_total"]["costo_merma"], ZERO)

    def test_html_exposes_point_authority_snapshots_and_conversion_source(self):
        context, _ = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(
                    receta_id=self.parent.id,
                    conversion_in=Decimal("2"),
                    conversion_origin="POINT",
                    status="COINCIDE",
                )
            )
        )
        rendered = self._render(context)
        self.assertIn("Conversiones Point: PointConversionLine", rendered)
        self.assertIn("Autoridad Point: Verificada", rendered)
        self.assertIn("Cierres Point · Inicial: 31/07/2026 · Final: 31/08/2026", rendered)
        self.assertIn("Origen: Point", rendered)
        self.assertNotIn('title="Origen de conversión:', rendered)

    def test_period_selector_applies_immediately_and_names_the_loaded_period(self):
        context, _ = self._context(
            canonical_balance(MonthlyPointBalanceRow(receta_id=self.parent.id, status="COINCIDE"))
        )

        rendered = self._render(context)

        self.assertIn('data-period-autosubmit', rendered)
        self.assertIn('Periodo mostrado: Agosto 2026', rendered)
        self.assertIn('requestSubmit()', rendered)
        self.assertIn('styles.css?v=20260903-report-point-v1', rendered)

    def test_partial_point_data_is_summarized_without_exposing_technical_wall(self):
        sources = canonical_balance().sources
        sources["opening_snapshot"].update({
            "authoritative": False,
            "source_present": False,
            "effective_date": None,
            "target_date": date(2026, 7, 31),
            "authority_issues": ("SNAPSHOT_SYNC_JOB_MISSING",),
        })
        sources["closing_snapshot"].update({
            "authoritative": False,
            "source_present": True,
            "effective_date": date(2026, 8, 31),
            "target_date": date(2026, 8, 31),
            "authority_issues": ("SNAPSHOT_SYNC_JOB_MISSING",),
        })
        balance = canonical_balance(
            MonthlyPointBalanceRow(
                receta_id=self.parent.id,
                closing_point=Decimal("22"),
                status="REVISAR_FUENTE",
            ),
            sources=sources,
            warnings=("No hay cierre Point acreditado para inicial del 2026-07-31.",),
            issues=("MONTH_SOURCE_INCOMPLETE",),
        )
        balance.effective_snapshot_dates = {"opening": None, "closing": date(2026, 8, 31)}
        context, _ = self._context(balance)

        rendered = self._render(context)

        self.assertIn("Información parcial", rendered)
        self.assertIn("Falta recuperar el cierre Point del 31/07/2026", rendered)
        self.assertIn("El cierre Point del 31/08/2026 sí está disponible", rendered)
        self.assertNotIn('class="badge bg-warning">Snapshot inicial Point:', rendered)
        self.assertIn("Ver diagnóstico técnico", rendered)
        self.assertIn("falta job Point del mes", rendered)

    def test_current_month_does_not_report_the_future_closing_as_missing(self):
        sources = canonical_balance().sources
        sources["opening_snapshot"].update({
            "effective_date": date(2026, 8, 31),
            "target_date": date(2026, 8, 31),
        })
        sources["closing_snapshot"].update({
            "authoritative": False,
            "source_present": False,
            "effective_date": None,
            "target_date": date(2026, 9, 30),
            "not_due": True,
            "authority_issues": (),
        })
        balance = canonical_balance(
            MonthlyPointBalanceRow(
                receta_id=self.parent.id,
                opening_point=Decimal("10"),
                production=Decimal("3"),
                sales=Decimal("2"),
                calculated_closing=Decimal("11"),
                closing_point=None,
                difference_point=None,
                status="EN_CURSO",
            ),
            sources=sources,
        )
        balance.effective_snapshot_dates = {"opening": date(2026, 8, 31), "closing": None}

        with patch("reportes.views_produccion.timezone.localdate", return_value=date(2026, 9, 4)):
            context, _ = self._context(balance, period="2026-09")
            rendered = self._render(context)

        self.assertIn("Periodo en curso", rendered)
        self.assertIn("Final: Al cierre del 30/09/2026", rendered)
        self.assertIn("movimientos Point del 01/09/2026 al 03/09/2026", rendered)
        self.assertNotIn("Falta recuperar el cierre Point final del 30/09/2026", rendered)
        self.assertNotIn("Final: Pendiente", rendered)
        self.assertEqual(context["groups"][0]["rows"][0]["estado_inventario"], "Periodo en curso")

    def test_bridge_history_or_non_authoritative_sales_never_verify_point_authority(self):
        sources = canonical_balance().sources
        sources["sales"] = {
            "source": "PointDailySale",
            "selected_source": "PointDailySale",
            "mode": "BRIDGE_HISTORY",
            "authoritative": False,
        }
        context, _ = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(receta_id=self.parent.id, status="COINCIDE"),
                sources=sources,
            )
        )
        rendered = self._render(context)
        self.assertIn("Autoridad Point: Revisar fuentes", rendered)
        self.assertIn("Ventas: BRIDGE_HISTORY", rendered)
        self.assertNotIn("Autoridad Point: Verificada", rendered)

    def test_required_source_present_false_blocks_authority_and_uses_honest_fallback_labels(self):
        sources = canonical_balance().sources
        sources["opening_snapshot"]["source_present"] = False
        sources["closing_snapshot"]["source_present"] = False
        sources["sales"]["source_present"] = False
        sources["production"] = {
            "source": "FactProduccionDiaria",
            "authoritative": True,
            "source_present": False,
        }
        sources["waste"] = {
            "source": "MermaMensualSucursal",
            "authoritative": True,
            "source_present": False,
        }
        context, _ = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(receta_id=self.parent.id, status="COINCIDE"),
                sources=sources,
            )
        )
        rendered = self._render(context)
        self.assertIn("Autoridad Point: Revisar fuentes", rendered)
        self.assertIn("Snapshot inicial Point: PointInventorySnapshot no disponible", rendered)
        self.assertIn("Snapshot final Point: PointInventorySnapshot no disponible", rendered)
        self.assertIn("Ventas: PointDailySale no disponible", rendered)
        self.assertIn("Producción: FactProduccionDiaria no disponible", rendered)
        self.assertIn("Merma: MermaMensualSucursal no disponible", rendered)
        self.assertIn("Producción: FactProduccionDiaria", rendered)
        self.assertIn("Merma: MermaMensualSucursal", rendered)
        self.assertNotIn("Producción Point: FactProduccionDiaria", rendered)
        self.assertNotIn("Merma Point: MermaMensualSucursal", rendered)

    def test_movement_authority_reasons_are_preserved_and_explained(self):
        sources = canonical_balance().sources
        sources["production"] = {
            "source": "PointProductionLine",
            "authoritative": False,
            "source_present": True,
            "job_status": "PARTIAL",
            "authority_issues": (
                "PRODUCTION_SYNC_JOB_PARTIAL",
                "PRODUCTION_SYNC_RANGE_INCOMPLETE",
            ),
        }

        context, _ = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(receta_id=self.parent.id, status="REVISAR_FUENTE"),
                sources=sources,
            )
        )
        rendered = self._render(context)

        self.assertEqual(
            context["fuentes"]["canonical"]["production"]["authority_issues"],
            ("PRODUCTION_SYNC_JOB_PARTIAL", "PRODUCTION_SYNC_RANGE_INCOMPLETE"),
        )
        self.assertIn("Producción: job Point parcial", rendered)
        self.assertIn("Producción: rango mensual incompleto", rendered)

    def test_available_periods_include_months_with_only_point_sales_or_conversions(self):
        branch = PointBranch.objects.create(external_id="period-branch", name="Sucursal periodos")
        product = PointProduct.objects.create(external_id="period-product", name="Producto periodos")
        PointDailySale.objects.create(
            branch=branch,
            product=product,
            sale_date=date(2026, 7, 15),
            quantity=Decimal("1"),
        )
        PointConversionLine.objects.create(
            branch=branch,
            movement_external_id="period-conversion",
            source_hash="period-conversion-hash",
            movement_at=timezone.make_aware(datetime(2026, 6, 20, 10, 0)),
            item_name="Conversión periodos",
            quantity=Decimal("1"),
        )

        periodos = ProducidoVsVendidoMermaView()._available_periods(selected="2026-01")

        self.assertIn("2026-07", periodos)
        self.assertIn("2026-06", periodos)

    def test_template_explains_difference_sign_and_only_shows_conversion_origin_for_activity(self):
        context, _ = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(receta_id=self.parent.id, status="COINCIDE"),
                MonthlyPointBalanceRow(
                    receta_id=self.slice.id,
                    conversion_in=Decimal("1"),
                    conversion_origin="EQUIVALENCIA_CONFIGURADA",
                    status="COINCIDE",
                ),
            )
        )
        rendered = self._render(context)
        slice_row = re.search(
            r'<tr>\s*<td class="cell-wrap">Rebanada canónica</td>(.*?)</tr>',
            rendered,
            flags=re.DOTALL,
        )
        self.assertIsNotNone(slice_row)
        conversion_cells = re.findall(
            r'<td class="text-end">(.*?)</td>',
            slice_row.group(1),
            flags=re.DOTALL,
        )
        self.assertIn("positivo = Point reporta más", rendered)
        self.assertIn("negativo = Point reporta menos", rendered)
        self.assertIn("cualquier valor distinto de cero requiere revisión", rendered)
        self.assertIn(
            "La Dif. Point no identifica ni explica por sí sola la causa y no atribuye responsabilidad.",
            rendered,
        )
        self.assertIn("Origen: equivalencia configurada", rendered)
        self.assertIn("Origen: equivalencia configurada", conversion_cells[4])
        self.assertNotIn("Origen: equivalencia configurada", conversion_cells[5])
        self.assertNotIn("Origen: Sin dato", rendered)

    def test_sources_and_issues_become_concise_traceable_banners(self):
        context, _ = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(receta_id=self.parent.id, status="REVISAR_FUENTE"),
                warnings=("Snapshot final fuera de tolerancia.",),
                issues=("opening_snapshot_missing",),
            )
        )
        self.assertIn("canonical", context["fuentes"])
        self.assertIn("selected_source", context["fuentes"]["canonical"]["sales"])
        self.assertTrue(any("Fuente Point" in banner for banner in context["banners"]))
        self.assertTrue(any("Snapshot" in banner for banner in context["banners"]))

    def test_july_and_august_use_the_same_local_only_service_contract(self):
        with patch("reportes.views_produccion.MonthlyPointProductBalanceService") as service:
            service.return_value.build.return_value = canonical_balance(
                MonthlyPointBalanceRow(receta_id=self.parent.id, status="REVISAR_FUENTE")
            )
            view = ProducidoVsVendidoMermaView()
            view._build_context(self.factory.get("/reportes/produccion/", {"periodo": "2026-07"}))
            view._build_context(self.factory.get("/reportes/produccion/", {"periodo": "2026-08"}))
        self.assertEqual(
            service.return_value.build.call_args_list,
            [call(month="2026-07"), call(month="2026-08")],
        )
        self.assertEqual(service.call_args_list, [call(), call()])
        self.assertFalse(service.return_value.build.call_args.kwargs.get("refresh_official_sales", False))

    def test_template_uses_point_language_and_formula_without_physical_inventory_claims(self):
        template = Path("reportes/templates/reportes/producido_vs_vendido.html").read_text()
        self.assertIn("Ini. Point", template)
        self.assertIn("Saldo calc.", template)
        self.assertIn("Fin. Point", template)
        self.assertIn("Dif. Point", template)
        self.assertIn(
            "Saldo calculado = inicial Point + producción + conversión de entrada − venta − merma − conversión de salida.",
            template,
        )
        self.assertIn("Dif. Point = final Point − saldo calculado.", template)
        self.assertNotIn("Sobrante físico", template)
        self.assertNotIn("Faltante no explicado", template)
        self.assertNotIn("Inventario físico", template)

    def test_csv_export_uses_canonical_point_labels_sign_status_and_sin_dato(self):
        context, _ = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(
                    receta_id=self.parent.id,
                    opening_point=Decimal("10"),
                    calculated_closing=Decimal("9"),
                    closing_point=Decimal("11"),
                    difference_point=Decimal("2"),
                    status="POINT_MAYOR",
                ),
                MonthlyPointBalanceRow(
                    receta_id=self.slice.id,
                    opening_point=None,
                    calculated_closing=None,
                    closing_point=None,
                    difference_point=None,
                    status="REVISAR_FUENTE",
                ),
            )
        )

        response = ProducidoVsVendidoMermaView()._export_csv(context)
        body = response.content.decode("utf-8")

        self.assertIn("Saldo calculado", body)
        self.assertIn("Fin. Point", body)
        self.assertIn("Dif. Point", body)
        self.assertIn("Conv. entrada Point", body)
        self.assertIn("Conv. salida Point", body)
        self.assertIn("Point mayor", body)
        self.assertIn(",9,11,2,Point mayor", body)
        self.assertIn("Sin dato", body)
        self.assertNotIn("Físico", body)
        self.assertNotIn("Sobrante físico", body)
        self.assertNotIn("Faltante no explicado", body)

    def test_csv_export_includes_exact_conversion_provenance(self):
        response = ProducidoVsVendidoMermaView()._export_csv(self._conversion_origin_context())
        body = response.content.decode("utf-8")

        self.assertIn("Procedencia conversión", body)
        for label in ("Point", "equivalencia configurada", "mixta", "sin resolver"):
            self.assertIn(label, body)

    def test_xlsx_export_uses_canonical_point_headers_and_preserves_missing_values(self):
        context, _ = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(
                    receta_id=self.parent.id,
                    opening_point=None,
                    calculated_closing=None,
                    closing_point=None,
                    difference_point=None,
                    status="REVISAR_FUENTE",
                )
            )
        )

        response = ProducidoVsVendidoMermaView()._export_xlsx(context)
        workbook = load_workbook(BytesIO(response.content), data_only=True, read_only=True)
        sheet = workbook["Producido vs Vendido"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=5, max_row=5))]
        detail = next(sheet.iter_rows(min_row=7, max_row=7, values_only=True))

        self.assertIn("Saldo calculado", headers)
        self.assertIn("Fin. Point", headers)
        self.assertIn("Dif. Point", headers)
        self.assertIn("Conv. entrada Point", headers)
        self.assertIn("Conv. salida Point", headers)
        self.assertNotIn("Inv. físico registrado", headers)
        self.assertEqual(detail[headers.index("Saldo calculado")], "Sin dato")
        self.assertEqual(detail[headers.index("Fin. Point")], "Sin dato")
        self.assertEqual(detail[headers.index("Dif. Point")], "Sin dato")

    def test_xlsx_export_includes_exact_conversion_provenance(self):
        response = ProducidoVsVendidoMermaView()._export_xlsx(self._conversion_origin_context())
        workbook = load_workbook(BytesIO(response.content), data_only=True, read_only=True)
        sheet = workbook["Producido vs Vendido"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=5, max_row=5))]
        provenance_index = headers.index("Procedencia conversión")
        values = {
            row[provenance_index]
            for row in sheet.iter_rows(min_row=6, values_only=True)
        }

        self.assertTrue({"Point", "equivalencia configurada", "mixta", "sin resolver"}.issubset(values))

    def test_pdf_export_uses_canonical_point_labels_and_neutral_status(self):
        context, _ = self._context(
            canonical_balance(
                MonthlyPointBalanceRow(
                    receta_id=self.parent.id,
                    opening_point=Decimal("10"),
                    calculated_closing=Decimal("9"),
                    closing_point=Decimal("11"),
                    difference_point=Decimal("2"),
                    status="POINT_MAYOR",
                )
            )
        )

        response = ProducidoVsVendidoMermaView()._export_pdf(context)
        body = response.content.decode("latin-1")

        self.assertIn("Ini. Point 10", body)
        self.assertIn("Saldo calculado 9", body)
        self.assertIn("Fin. Point 11", body)
        self.assertIn("Dif. Point 2", body)
        self.assertIn("Conv. entrada Point", body)
        self.assertIn("Conv. salida Point", body)
        self.assertIn("positivo = Point mayor", body)
        self.assertLess(body.index("Ini. Point 10"), body.index("Saldo calculado 9"))
        self.assertIn("Estado Point mayor", body)
        self.assertNotIn("Físico", body)
        self.assertNotIn("Sobrante físico", body)
        self.assertNotIn("Faltante no explicado", body)

    def test_pdf_export_includes_exact_conversion_provenance(self):
        response = ProducidoVsVendidoMermaView()._export_pdf(self._conversion_origin_context())
        body = response.content.decode("latin-1")

        for label in ("Point", "equivalencia configurada", "mixta", "sin resolver"):
            self.assertIn(f"Procedencia conversion: {label}", body)
