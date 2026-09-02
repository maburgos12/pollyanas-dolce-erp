from __future__ import annotations

import csv
import textwrap
from calendar import monthrange
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.db.models.functions import TruncMonth
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.utils import timezone
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from control.models import MermaMensualSucursal
from core.access import can_view_reportes
from pos_bridge.models import (
    PointConversionLine,
    PointDailySale,
    PointInventorySnapshot,
    PointProductionLine,
    PointSalesDailyProductFact,
    PointWasteLine,
)
from pos_bridge.services.monthly_product_balance_service import MonthlyPointProductBalanceService
from recetas.models import ProductoMonthClosure, Receta
from recetas.utils.derived_product_presentations import get_total_cost_map
from reportes.models import FactProduccionDiaria


ZERO = Decimal("0")

CATEGORY_ORDER = [
    "Pastel Mini",
    "Pastel Chico",
    "Pastel Mediano",
    "Pastel Grande",
    "Pastel Individual",
    "Individual",
    "Media Plancha",
    "Rosca",
    "Rebanada",
    "Pay Mediano",
    "Pay Grande",
    "Vaso Preparado Mini",
    "Vasos Mini",
    "Vasos Grande",
    "Vasos Preparados Grande",
    "Bollo",
    "Cheesecake",
    "Empanadas",
    "Galletas",
    "Otros postres",
    "Café",
]
CATEGORY_ORDER_INDEX = {name.lower(): index for index, name in enumerate(CATEGORY_ORDER)}

PRODUCTION_EXPORT_COLUMNS = [
    ("categoria", "Categoría", "text"),
    ("receta", "Receta", "text"),
    ("vendido", "Vendido", "number"),
    ("producido", "Producido", "number"),
    ("dif", "Dif. operativa", "number"),
    ("merma_reportada", "Merma", "number"),
    ("conversion_entrada", "Conv.", "number"),
    ("enteros_equivalentes", "Eq.", "number"),
    ("costo_merma", "Costo merma", "currency"),
    ("pct_merma", "% merma", "percent"),
    ("inventario_inicial", "Ini. Point", "number"),
    ("inventario_final_teorico", "Saldo calculado", "number"),
    ("inventario_final_point_total", "Fin. Point", "number"),
    ("diferencia_inventario", "Dif. Point", "number"),
    ("estado_inventario", "Estado", "text"),
]

PDF_EXPORT_COLUMNS = [
    ("vendido", "Vta."),
    ("producido", "Prod."),
    ("dif", "Dif."),
    ("merma_reportada", "Merma"),
    ("pct_merma", "%"),
    ("inventario_final_teorico", "Saldo calculado"),
    ("inventario_final_point_total", "Fin. Point"),
    ("diferencia_inventario", "Dif. Point"),
]


@dataclass(frozen=True)
class PeriodSelection:
    month_start: date
    month_end: date

    @property
    def value(self) -> str:
        return self.month_start.strftime("%Y-%m")

    @property
    def label(self) -> str:
        return self.month_start.strftime("%B %Y").title()


def _parse_period(raw_value: str | None) -> PeriodSelection:
    today = timezone.localdate()
    fallback = date(today.year, today.month, 1)
    if raw_value:
        try:
            year_raw, month_raw = raw_value.split("-", 1)
            selected = date(int(year_raw), int(month_raw), 1)
        except (TypeError, ValueError):
            selected = fallback
    else:
        selected = fallback
    last_day = monthrange(selected.year, selected.month)[1]
    return PeriodSelection(month_start=selected, month_end=date(selected.year, selected.month, last_day))


def _parse_int(raw_value: str | None) -> int | None:
    try:
        value = int(raw_value or 0)
    except (TypeError, ValueError):
        return None
    return value or None


def _decimal(value: Any) -> Decimal:
    if value is None:
        return ZERO
    return Decimal(str(value or 0))


def _category_label(value: str | None) -> str:
    return (value or "").strip() or "Sin categoría"


def _category_sort_key(value: str | None) -> tuple[int, str]:
    label = _category_label(value)
    return (CATEGORY_ORDER_INDEX.get(label.lower(), len(CATEGORY_ORDER)), label.lower())


def _is_production_reference_category(value: str | None) -> bool:
    return _category_label(value).lower() == "rebanada"


def _sum_or_none(rows: list[dict[str, Any]], key: str) -> Decimal | None:
    if not rows or any(row.get(key) is None for row in rows):
        return None
    return sum((row[key] for row in rows), ZERO)


def _filename_period(context: dict[str, Any]) -> str:
    return str(context.get("selected_period") or timezone.localdate().strftime("%Y-%m"))


def _decimal_or_none(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except Exception:
        return None


def _format_decimal(value: Any, *, places: int = 2, trim: bool = True) -> str:
    decimal_value = _decimal_or_none(value)
    if decimal_value is None:
        return ""
    formatted = f"{decimal_value:,.{places}f}"
    if trim and "." in formatted:
        formatted = formatted.rstrip("0").rstrip(".")
    return formatted


def _export_raw_value(row: dict[str, Any], key: str) -> Any:
    if key == "dif" and row.get("produccion_referencia"):
        return "Referencia"
    return row.get(key)


def _export_display_value(row: dict[str, Any], key: str, kind: str) -> str:
    value = _export_raw_value(row, key)
    if value is None or value == "":
        return "Sin dato"
    if kind == "currency":
        return f"${_format_decimal(value, places=2, trim=False)}"
    if kind == "percent":
        return f"{_format_decimal(value, places=2, trim=False)}%"
    if kind == "number":
        if isinstance(value, str):
            return value
        return _format_decimal(value, places=2)
    return str(value)


def _export_rows(context: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for group in context.get("groups") or []:
        subtotal = dict(group.get("total") or {})
        subtotal.update(
            {
                "_row_type": "subtotal",
                "categoria": group.get("categoria") or group.get("familia") or "",
                "receta": "Subtotal",
                "estado_inventario": "Subtotal",
            }
        )
        rows.append(subtotal)
        for detail in group.get("rows") or []:
            row = dict(detail)
            row["_row_type"] = "detail"
            rows.append(row)

    total = dict(context.get("grand_total") or {})
    total.update(
        {
            "_row_type": "total",
            "categoria": "Gran total",
            "receta": "Total",
            "estado_inventario": "Total",
        }
    )
    rows.append(total)
    return rows


def _pdf_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _pdf_bytes(*, title: str, lines: list[str]) -> bytes:
    page_width = 792
    page_height = 612
    line_height = 12
    max_lines = 43
    pages = [lines[index : index + max_lines] for index in range(0, len(lines), max_lines)] or [[]]
    font_id = 3 + (len(pages) * 2)
    page_ids = [3 + (index * 2) for index in range(len(pages))]

    objects: list[bytes] = [
        b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj",
        (
            "2 0 obj << /Type /Pages /Count {count} /Kids [{kids}] >> endobj".format(
                count=len(pages),
                kids=" ".join(f"{page_id} 0 R" for page_id in page_ids),
            )
        ).encode(),
    ]

    for index, page_lines in enumerate(pages):
        page_id = page_ids[index]
        content_id = page_id + 1
        content_parts = ["BT", "/F1 8 Tf", f"{line_height} TL", "36 560 Td"]
        page_title = title if index == 0 else f"{title} - pág. {index + 1}"
        for raw in [page_title, "", *page_lines]:
            content_parts.append(f"({_pdf_escape(raw)}) Tj")
            content_parts.append("T*")
        content_parts.append("ET")
        content = "\n".join(content_parts).encode("latin-1", errors="replace")
        objects.append(
            (
                f"{page_id} 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width} {page_height}] "
                f"/Resources << /Font << /F1 {font_id} 0 R >> >> /Contents {content_id} 0 R >> endobj"
            ).encode()
        )
        objects.append(
            b"%d 0 obj << /Length " % content_id
            + str(len(content)).encode()
            + b" >> stream\n"
            + content
            + b"\nendstream endobj"
        )

    objects.append(f"{font_id} 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj".encode())

    output = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(len(output))
        output.extend(obj)
        output.extend(b"\n")
    xref_pos = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode())
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode())
    output.extend(
        f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_pos}\n%%EOF".encode()
    )
    return bytes(output)


class ProducidoVsVendidoMermaView(LoginRequiredMixin, TemplateView):
    template_name = "reportes/producido_vs_vendido.html"

    def dispatch(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        if request.user.is_authenticated and not can_view_reportes(request.user):
            raise PermissionDenied("No tienes permisos para ver Reportes.")
        return super().dispatch(request, *args, **kwargs)

    def get(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        context = self._build_context(request)
        if request.resolver_match and request.resolver_match.url_name == "producido_vs_vendido_data":
            return JsonResponse(
                {
                    "periodo": context["selected_period"],
                    "fuentes": context["fuentes"],
                    "rows": context["json_rows"],
                    "totals": context["grand_total"],
                }
            )
        export_format = (request.GET.get("export") or "").strip().lower()
        if export_format == "csv":
            return self._export_csv(context)
        if export_format == "xlsx":
            return self._export_xlsx(context)
        if export_format == "pdf":
            return self._export_pdf(context)
        return self.render_to_response(context)

    def _export_csv(self, context: dict[str, Any]) -> HttpResponse:
        period = _filename_period(context)
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="producido_vs_vendido_{period}.csv"'
        writer = csv.writer(response)
        writer.writerow([label for _, label, _ in PRODUCTION_EXPORT_COLUMNS])
        for row in _export_rows(context):
            writer.writerow(
                [
                    _export_display_value(row, key, kind)
                    for key, _, kind in PRODUCTION_EXPORT_COLUMNS
                ]
            )
        return response

    def _export_xlsx(self, context: dict[str, Any]) -> HttpResponse:
        period = _filename_period(context)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Producido vs Vendido"

        title_font = Font(bold=True, size=14, color="7B1A48")
        header_fill = PatternFill("solid", fgColor="F5E6ED")
        header_font = Font(bold=True, color="7B1A48")
        subtotal_fill = PatternFill("solid", fgColor="F5E6ED")
        total_fill = PatternFill("solid", fgColor="3D0A24")
        white_bold = Font(bold=True, color="FFFFFF")

        sheet["A1"] = f"Producido vs Vendido - {context.get('selected_period_label') or period}"
        sheet["A1"].font = title_font
        sheet["A2"] = (
            f"Ventas: {context['fuentes']['ventas']['label']} | "
            f"Producción: {context['fuentes']['produccion']['label']} | "
            f"Merma: {context['fuentes']['merma']['label']} | "
            f"Inventario: {context['fuentes']['inventario']['label']}"
        )
        sheet["A3"] = f"Categoría: {context.get('selected_categoria') or 'Todas'}"

        header_row = 5
        for col_idx, (_, label, _) in enumerate(PRODUCTION_EXPORT_COLUMNS, start=1):
            cell = sheet.cell(row=header_row, column=col_idx, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, row in enumerate(_export_rows(context), start=header_row + 1):
            for col_idx, (key, _, kind) in enumerate(PRODUCTION_EXPORT_COLUMNS, start=1):
                raw_value = _export_raw_value(row, key)
                cell = sheet.cell(row=row_idx, column=col_idx)
                if raw_value is None or raw_value == "":
                    cell.value = "Sin dato"
                elif kind == "currency":
                    cell.value = _decimal_or_none(raw_value)
                    cell.number_format = '$#,##0.00'
                elif kind == "percent":
                    decimal_value = _decimal_or_none(raw_value)
                    cell.value = decimal_value / Decimal("100") if decimal_value is not None else None
                    cell.number_format = '0.00%'
                elif kind == "number" and not isinstance(raw_value, str):
                    cell.value = _decimal_or_none(raw_value)
                    cell.number_format = '#,##0.##'
                else:
                    cell.value = str(raw_value)
                cell.alignment = Alignment(
                    horizontal="left" if kind == "text" else "right",
                    vertical="center",
                    wrap_text=True,
                )

            if row.get("_row_type") == "subtotal":
                for cell in sheet[row_idx]:
                    cell.fill = subtotal_fill
                    cell.font = Font(bold=True, color="7B1A48")
            elif row.get("_row_type") == "total":
                for cell in sheet[row_idx]:
                    cell.fill = total_fill
                    cell.font = white_bold

        sheet.freeze_panes = "A6"
        widths = {
            "A": 22,
            "B": 34,
            "I": 15,
            "L": 16,
            "M": 18,
            "O": 20,
        }
        for col_idx in range(1, len(PRODUCTION_EXPORT_COLUMNS) + 1):
            letter = get_column_letter(col_idx)
            sheet.column_dimensions[letter].width = widths.get(letter, 12)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="producido_vs_vendido_{period}.xlsx"'
        return response

    def _export_pdf(self, context: dict[str, Any]) -> HttpResponse:
        period = _filename_period(context)
        lines = [
            f"Periodo: {context.get('selected_period_label') or period}",
            f"Categoria: {context.get('selected_categoria') or 'Todas'}",
            (
                f"Fuentes - Ventas: {context['fuentes']['ventas']['label']} | "
                f"Produccion: {context['fuentes']['produccion']['label']} | "
                f"Merma: {context['fuentes']['merma']['label']} | "
                f"Inventario: {context['fuentes']['inventario']['label']}"
            ),
            "",
        ]
        for row in _export_rows(context):
            prefix = "TOTAL" if row.get("_row_type") == "total" else "SUBTOTAL" if row.get("_row_type") == "subtotal" else "RECETA"
            name_line = f"{prefix}: {row.get('categoria') or ''} - {row.get('receta') or ''}".strip()
            lines.extend(textwrap.wrap(name_line, width=118) or [""])
            metrics = " | ".join(
                f"{label} {_export_display_value(row, key, 'percent' if key == 'pct_merma' else 'number') or '-'}"
                for key, label in PDF_EXPORT_COLUMNS
            )
            status = _export_display_value(row, "estado_inventario", "text") or "-"
            lines.extend(textwrap.wrap(f"{metrics} | Estado {status}", width=118, subsequent_indent="  "))
            costo = _export_display_value(row, "costo_merma", "currency") or "$0.00"
            lines.append(f"Costo merma: {costo}")
            lines.append("")

        pdf = _pdf_bytes(title="Producido vs Vendido", lines=lines)
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="producido_vs_vendido_{period}.pdf"'
        return response

    def _build_context(self, request: HttpRequest) -> dict[str, Any]:
        period = _parse_period(request.GET.get("periodo") or request.GET.get("period"))
        categoria = (request.GET.get("categoria") or request.GET.get("familia") or "").strip()

        # This report is deliberately read-only: the canonical balance service defaults
        # to local facts/snapshots and never requests a Point refresh from this path.
        balance = MonthlyPointProductBalanceService().build(month=period.value)
        recipe_ids = set(balance.rows)

        recipes_qs = Receta.objects.filter(
            id__in=recipe_ids,
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        if categoria:
            recipes_qs = recipes_qs.filter(categoria=categoria)
        recipes = sorted(
            list(recipes_qs),
            key=lambda recipe: (_category_sort_key(recipe.categoria), recipe.nombre.lower()),
        )

        cost_map = get_total_cost_map([recipe.id for recipe in recipes])
        rows = [
            self._build_row(recipe, balance.rows[recipe.id], cost_map, balance.sources)
            for recipe in recipes
        ]
        groups, grand_total = self._group_rows(rows)
        fuentes = self._canonical_sources(balance)
        banners = self._canonical_banners(balance, fuentes)
        periodos = self._available_periods(selected=period.value)

        return {
            "module_tabs": self._module_tabs("producido_vs_vendido"),
            "selected_period": period.value,
            "selected_period_label": period.label,
            "periodos": periodos,
            "selected_categoria": categoria,
            "selected_familia": categoria,
            "categorias": self._categories(),
            "familias": self._categories(),
            "groups": groups,
            "grand_total": grand_total,
            "json_rows": [row["json"] for row in rows],
            "fuentes": fuentes,
            "banners": banners,
            "source_dates": balance.effective_snapshot_dates,
        }

    def _build_row(
        self,
        recipe: Receta,
        balance_row,
        cost_map: dict[int, Decimal],
        sources: dict[str, Any],
    ) -> dict[str, Any]:
        vendido = self._movement_value(balance_row.sales, sources.get("sales"))
        producido = self._movement_value(balance_row.production, sources.get("production"))
        merma_reportada = self._movement_value(balance_row.waste, sources.get("waste"))
        categoria = _category_label(recipe.categoria)
        produccion_referencia = not bool(recipe.pasa_modulo_produccion)
        dif = None
        dif_referencia = None
        if producido is not None and vendido is not None:
            raw_dif = producido - vendido
            if produccion_referencia:
                dif_referencia = raw_dif
            else:
                dif = raw_dif
        costo_unitario = cost_map.get(recipe.id, ZERO)
        costo_merma = merma_reportada * costo_unitario if merma_reportada is not None and costo_unitario else None
        pct_merma = None
        if merma_reportada is not None and vendido and vendido > ZERO:
            pct_merma = (merma_reportada / vendido) * Decimal("100")
        snapshots_authoritative = self._snapshots_are_authoritative(sources)
        row = {
            "receta_id": recipe.id,
            "receta": recipe.nombre,
            "categoria": categoria,
            "familia": categoria,
            "vendido": vendido,
            "producido": producido,
            "dif": dif,
            "dif_referencia": dif_referencia,
            "produccion_referencia": produccion_referencia,
            "merma_reportada": merma_reportada,
            "costo_merma": costo_merma,
            "pct_merma": pct_merma,
            "convertido": balance_row.conversion_in,
            "enteros_equivalentes": balance_row.conversion_out,
            "conversion_entrada": balance_row.conversion_in,
            "conversion_salida": balance_row.conversion_out,
            "conversion_provenance": balance_row.conversion_origin or "Sin dato",
            "conversion_provenance_label": self._conversion_provenance_label(balance_row.conversion_origin),
            "inventario_inicial": balance_row.opening_point if snapshots_authoritative else None,
            "inventario_final_teorico": balance_row.calculated_closing if snapshots_authoritative else None,
            "inventario_final_point_total": balance_row.closing_point if snapshots_authoritative else None,
            "diferencia_inventario": balance_row.difference_point if snapshots_authoritative else None,
            "estado_inventario": self._point_status_label(balance_row.status),
        }
        row["json"] = {
            key: (str(value) if isinstance(value, Decimal) else value)
            for key, value in row.items()
            if key != "json"
        }
        return row

    @staticmethod
    def _movement_value(value: Decimal, source: dict[str, Any] | None) -> Decimal | None:
        if source and source.get("source_present") is False:
            return None
        return value

    @staticmethod
    def _snapshots_are_authoritative(sources: dict[str, Any]) -> bool:
        return all(bool((sources.get(key) or {}).get("authoritative")) for key in ("opening_snapshot", "closing_snapshot"))

    @staticmethod
    def _point_status_label(status: str) -> str:
        return {
            "COINCIDE": "Coincide",
            "POINT_MAYOR": "Point mayor",
            "POINT_MENOR": "Point menor",
            "REVISAR_FUENTE": "Revisar fuente",
        }.get(status, "Revisar fuente")

    @staticmethod
    def _conversion_provenance_label(origin: str) -> str:
        return {
            "POINT": "Point",
            "EQUIVALENCIA_CONFIGURADA": "equivalencia configurada",
            "MIXED": "orígenes mixtos",
            "UNRESOLVED": "Revisar fuente",
        }.get(origin, "Sin dato")

    @staticmethod
    def _source_descriptor(source: dict[str, Any] | None) -> dict[str, Any]:
        source = dict(source or {})
        return {
            "source": source.get("source") or "Sin dato",
            "selected_source": source.get("selected_source") or source.get("source") or "Sin dato",
            "mode": source.get("mode") or source.get("configured_source_mode") or "Sin dato",
            "authoritative": source.get("authoritative"),
            "source_present": source.get("source_present"),
            "effective_date": source.get("effective_date"),
            "coverage": source.get("applied_coverage_key_count"),
            "target_date": source.get("target_date"),
            "selected_dates": source.get("selected_dates") or (),
            "fallback_used": bool(source.get("fallback_used")),
        }

    def _canonical_sources(self, balance) -> dict[str, Any]:
        canonical = {
            "opening": self._source_descriptor(balance.sources.get("opening_snapshot")),
            "closing": self._source_descriptor(balance.sources.get("closing_snapshot")),
            "production": self._source_descriptor(balance.sources.get("production")),
            "sales": self._source_descriptor(balance.sources.get("sales")),
            "waste": self._source_descriptor(balance.sources.get("waste")),
            "conversions": self._source_descriptor(balance.sources.get("conversions")),
            "snapshot_dates": dict(balance.effective_snapshot_dates),
        }
        canonical["authority"] = self._canonical_authority(balance, canonical)
        return {
            "ventas": {"label": canonical["sales"]["selected_source"], **canonical["sales"]},
            "produccion": {"label": canonical["production"]["source"], **canonical["production"]},
            "merma": {"label": canonical["waste"]["source"], **canonical["waste"]},
            "inventario": {"label": "Snapshots Point", **canonical["closing"]},
            "canonical": canonical,
        }

    @staticmethod
    def _canonical_authority(balance, canonical: dict[str, Any]) -> dict[str, Any]:
        reasons: list[str] = []
        for key, label in (
            ("opening", "Snapshot inicial Point"),
            ("closing", "Snapshot final Point"),
            ("sales", "Ventas"),
            ("production", "Producción"),
            ("waste", "Merma"),
        ):
            source = canonical[key]
            if source["source_present"] is False:
                reasons.append(f"{label}: {source['selected_source']} no disponible")
            if key in {"opening", "closing", "sales"} and not source["authoritative"]:
                reasons.append(f"{label}: {source['selected_source']} sin autoridad")

        sales = canonical["sales"]
        if sales["mode"] == "BRIDGE_HISTORY":
            reasons.append("Ventas: BRIDGE_HISTORY")
        if getattr(balance, "issues", ()):
            reasons.append("Incidencias canónicas pendientes")
        if getattr(balance, "unresolved_movements", ()) or getattr(balance, "unresolved_conversions", ()):
            reasons.append("Movimientos Point pendientes de resolver")
        if any(row.status == "REVISAR_FUENTE" for row in balance.rows.values()):
            reasons.append("Filas Point en revisión")
        return {
            "verified": not reasons,
            "label": "Verificada" if not reasons else "Revisar fuentes",
            "reason": " · ".join(reasons),
        }

    def _canonical_banners(self, balance, fuentes: dict[str, Any]) -> list[str]:
        sales = fuentes["canonical"]["sales"]
        opening = fuentes["canonical"]["opening"]
        closing = fuentes["canonical"]["closing"]
        banners = [
            "Fuente Point: ventas {sales}; inicial {opening}; final {closing}.".format(
                sales=sales["selected_source"],
                opening=opening["source"],
                closing=closing["source"],
            )
        ]
        if not self._snapshots_are_authoritative(balance.sources):
            banners.append("Snapshots Point no autoritativos: los saldos y su diferencia se muestran como Sin dato.")
        banners.extend(str(warning) for warning in balance.warnings)
        banners.extend(str(issue) for issue in balance.issues)
        return list(dict.fromkeys(banners))

    def _available_periods(self, *, selected: str) -> list[str]:
        months = {selected}
        for model, field in (
            (ProductoMonthClosure, "month_start"),
            (FactProduccionDiaria, "fecha"),
            (PointProductionLine, "production_date"),
            (PointSalesDailyProductFact, "sale_date"),
            (PointDailySale, "sale_date"),
            (PointConversionLine, "movement_at"),
            (PointWasteLine, "movement_at"),
            (MermaMensualSucursal, "periodo"),
            (PointInventorySnapshot, "captured_at"),
        ):
            values = (
                model.objects.annotate(report_month=TruncMonth(field))
                .order_by()
                .values_list("report_month", flat=True)
                .distinct()
            )
            for value in values:
                if value:
                    months.add(value.strftime("%Y-%m"))
        return sorted(months, reverse=True)

    def _group_rows(self, rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            grouped[row["categoria"]].append(row)

        groups = []
        grand_rows: list[dict[str, Any]] = []
        for category in sorted(grouped, key=_category_sort_key):
            family_rows = grouped[category]
            grand_rows.extend(family_rows)
            groups.append(
                {
                    "categoria": category,
                    "familia": category,
                    "rows": family_rows,
                    "total": self._totals(family_rows),
                }
            )
        return groups, self._totals(grand_rows)

    def _totals(self, rows: list[dict[str, Any]]) -> dict[str, Any]:
        totals = {
            "vendido": _sum_or_none(rows, "vendido"),
            "producido": _sum_or_none(rows, "producido"),
            "dif": self._difference_total(rows),
            "merma_reportada": _sum_or_none(rows, "merma_reportada"),
            "costo_merma": _sum_or_none(rows, "costo_merma"),
            "convertido": _sum_or_none(rows, "convertido"),
            "enteros_equivalentes": _sum_or_none(rows, "enteros_equivalentes"),
            "conversion_entrada": _sum_or_none(rows, "conversion_entrada"),
            "conversion_salida": _sum_or_none(rows, "conversion_salida"),
            "inventario_inicial": _sum_or_none(rows, "inventario_inicial"),
            "inventario_final_teorico": _sum_or_none(rows, "inventario_final_teorico"),
            "inventario_final_point_total": _sum_or_none(rows, "inventario_final_point_total"),
            "diferencia_inventario": _sum_or_none(rows, "diferencia_inventario"),
            "produccion_referencia": bool(rows) and all(row.get("produccion_referencia") for row in rows),
            "dif_referencia": _sum_or_none(rows, "dif_referencia"),
        }
        totals["pct_merma"] = (
            (totals["merma_reportada"] / totals["vendido"]) * Decimal("100")
            if totals["merma_reportada"] is not None and totals["vendido"] is not None and totals["vendido"] > ZERO
            else None
        )
        return totals

    @staticmethod
    def _difference_total(rows: list[dict[str, Any]]) -> Decimal | None:
        non_reference_rows = [row for row in rows if not row.get("produccion_referencia")]
        authoritative_values = [row["dif"] for row in non_reference_rows if row.get("dif") is not None]
        if not authoritative_values:
            return None
        return sum(authoritative_values, ZERO)

    def _categories(self) -> list[str]:
        values = (
            Receta.objects.filter(tipo=Receta.TIPO_PRODUCTO_FINAL)
            .exclude(categoria="")
            .values_list("categoria", flat=True)
            .distinct()
        )
        return sorted({_category_label(value) for value in values}, key=_category_sort_key)

    def _families(self) -> list[str]:
        return self._categories()

    def _module_tabs(self, active: str) -> list[dict[str, str | bool]]:
        tabs = [
            ("ventas", "/reportes/ventas/", "Ventas"),
            ("cierre_operativo", "/reportes/cierre-operativo/", "Cierre diario"),
            ("cierre_producto", "/reportes/cierre-producto/", "Cierre producto"),
            ("producido_vs_vendido", "/reportes/produccion/", "Producido vs Vendido"),
            ("financiero", "/reportes/financiero/", "Financiero"),
            ("mermas_devoluciones", "/reportes/mermas-devoluciones/", "Mermas y Devoluciones"),
            ("auditoria_insumos", "/reportes/auditoria-insumos/", "Auditoría Insumos"),
            ("proyeccion_produccion", "/reportes/proyeccion-produccion/", "Proyección Producción"),
            ("bi", "/reportes/bi/", "BI"),
        ]
        return [
            {"key": key, "url": url, "label": label, "active": key == active}
            for key, url, label in tabs
        ]
