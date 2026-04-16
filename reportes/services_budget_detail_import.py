from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
import re
from urllib.parse import unquote

from django.db import models, transaction
from openpyxl import load_workbook

from reportes.models import PresupuestoImport, PresupuestoLineaMensual
from reportes.services_budget_import import GeneralBudgetImportService, _safe_ratio, _sha256, _to_decimal


@dataclass
class TrustedBudgetDetailImportSummary:
    imports_created: int
    imports_updated: int
    lines_created: int
    lines_updated: int
    periods: list[str]
    sheets_imported: list[str]


class TrustedBudgetDetailImportService:
    SHEET_CONFIG = {
        "PRESUPUESTO ADMINISTRACIÓN 2026.xlsx": {
            "ADMON": "admin_recurrente",
        },
        "PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx": {
            "GUAMUCHIL": "branch_sales",
            "EL TUNEL": "branch_sales",
            "PLAZA NIO": "branch_sales",
            "CRUCERO": "branch_sales",
            "COLOSIO": "branch_sales",
            "GLORIAS": "branch_sales",
            "PAYAN": "branch_sales",
            "LEYVA": "branch_sales",
            "MATRIZ": "branch_sales",
        },
        "PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx": {
            "VENTAS": "payroll_area",
            "PRODUCCCION": "payroll_area",
            "LOGISTICA": "payroll_area",
            "ADMINISTRACION": "payroll_area",
        },
        "PRESUPUESTO PRODUCCIÓN 2026 AUTORIZADO.xlsx": {
            "PRESUPUESTO PRODUCCIÓN": "production_budget",
        },
        "PRESUPUESTO LOGISTICA 2026.xlsx": {
            "LOGÍSTICA": "logistics_budget",
        },
    }

    def __init__(self) -> None:
        self._general = GeneralBudgetImportService()

    def _iter_expected_files(self, folder: Path) -> list[Path]:
        paths: list[Path] = []
        for filename in self.SHEET_CONFIG:
            path = folder / filename
            if path.exists():
                paths.append(path)
        return paths

    def expected_sheets_for_file(self, filename: str) -> dict[str, str]:
        config = self.SHEET_CONFIG.get(filename)
        if config is None:
            raise ValueError(f"{filename} no está soportado como presupuesto detallado confiable.")
        return dict(config)

    def _import_sheet(
        self,
        *,
        path: Path,
        workbook_formula,
        workbook_values,
        sheet_name: str,
        kind: str,
    ) -> tuple[int, int, int, int, set[str]]:
        ws_formula = workbook_formula[sheet_name]
        ws_values = workbook_values[sheet_name]
        title = self._general._find_title(ws_formula)
        year = self._general._resolve_year(path, title)
        annual_cols, monthly_templates, data_start_row = self._general._parse_layout(ws_formula)
        monthly_groups = [(month_number, date(year, month_number, 1), cols) for month_number, cols in monthly_templates]
        if not monthly_groups:
            raise ValueError(f"No pude encontrar bloques mensuales válidos en {path.name} / {sheet_name}.")

        file_hash = _sha256(path)
        import_obj, created = PresupuestoImport.objects.update_or_create(
            tipo=PresupuestoImport.TIPO_DETALLE,
            fuente_nombre=path.name,
            sheet_name=sheet_name,
            defaults={
                "archivo_ruta": str(path),
                "archivo_hash": file_hash,
                "titulo": title,
                "metadata": {"year": year, "kind": kind},
            },
        )
        imports_created = int(created)
        imports_updated = int(not created)
        lines_created = 0
        lines_updated = 0
        periods: set[str] = set()

        started = False
        for row_index in range(data_start_row, ws_values.max_row + 1):
            raw_account = ws_values.cell(row_index, 1).value
            raw_concept = ws_values.cell(row_index, 2).value
            account_code = str(raw_account or "").strip()
            concept = str(raw_concept or "").strip()
            if not isinstance(raw_concept, str) and account_code:
                concept = account_code
                account_code = ""
            elif not concept and account_code:
                concept = account_code
                account_code = ""
            if not concept:
                if started:
                    continue
                continue
            started = True

            annual_budget = _to_decimal(
                ws_values.cell(row_index, annual_cols.get("budget", 0)).value if annual_cols.get("budget") else 0
            )
            annual_actual = _to_decimal(
                ws_values.cell(row_index, annual_cols.get("actual", 0)).value if annual_cols.get("actual") else 0
            )
            annual_variance = _to_decimal(
                ws_values.cell(row_index, annual_cols.get("variance", 0)).value if annual_cols.get("variance") else 0
            )
            if annual_variance == 0 and annual_budget > 0 and annual_actual > 0:
                annual_variance = _safe_ratio(annual_actual, annual_budget)

            for month_number, period_start, columns in monthly_groups:
                monthly_budget = _to_decimal(
                    ws_values.cell(row_index, columns.get("budget", 0)).value if columns.get("budget") else 0
                )
                monthly_actual = _to_decimal(
                    ws_values.cell(row_index, columns.get("actual", 0)).value if columns.get("actual") else 0
                )
                monthly_variance = _to_decimal(
                    ws_values.cell(row_index, columns.get("variance", 0)).value if columns.get("variance") else 0
                )
                if monthly_variance == 0 and monthly_budget > 0 and monthly_actual > 0:
                    monthly_variance = _safe_ratio(monthly_actual, monthly_budget)
                if (
                    annual_budget == 0
                    and annual_actual == 0
                    and monthly_budget == 0
                    and monthly_actual == 0
                    and monthly_variance == 0
                    and annual_variance == 0
                ):
                    continue

                external_key = (
                    f"{PresupuestoImport.TIPO_DETALLE}|{path.name}|{sheet_name}|{row_index}|{period_start.isoformat()}"
                )
                _, was_created = PresupuestoLineaMensual.objects.update_or_create(
                    external_key=external_key,
                    defaults={
                        "importacion": import_obj,
                        "period": period_start,
                        "account_code": account_code,
                        "concept": concept,
                        "annual_budget": annual_budget,
                        "annual_actual": annual_actual,
                        "annual_variance": annual_variance,
                        "monthly_budget": monthly_budget,
                        "monthly_actual": monthly_actual,
                        "monthly_variance": monthly_variance,
                        "row_index": row_index,
                        "metadata": {
                            "month_number": month_number,
                            "source_title": title,
                            "kind": kind,
                            "sheet_name": sheet_name,
                        },
                    },
                )
                lines_created += int(was_created)
                lines_updated += int(not was_created)
                periods.add(period_start.isoformat())

        return imports_created, imports_updated, lines_created, lines_updated, periods

    @transaction.atomic
    def import_folder(self, folder_path: str | Path) -> TrustedBudgetDetailImportSummary:
        folder = Path(folder_path).expanduser().resolve()
        if not folder.exists():
            raise FileNotFoundError(folder)

        imports_created = 0
        imports_updated = 0
        lines_created = 0
        lines_updated = 0
        periods: set[str] = set()
        sheets_imported: list[str] = []

        for path in self._iter_expected_files(folder):
            workbook_formula = load_workbook(path, read_only=True, data_only=False)
            workbook_values = load_workbook(path, read_only=True, data_only=True)
            for sheet_name, kind in self.SHEET_CONFIG[path.name].items():
                if sheet_name not in workbook_formula.sheetnames:
                    continue
                result = self._import_sheet(
                    path=path,
                    workbook_formula=workbook_formula,
                    workbook_values=workbook_values,
                    sheet_name=sheet_name,
                    kind=kind,
                )
                imports_created += result[0]
                imports_updated += result[1]
                lines_created += result[2]
                lines_updated += result[3]
                periods.update(result[4])
                sheets_imported.append(f"{path.name}::{sheet_name}")

        return TrustedBudgetDetailImportSummary(
            imports_created=imports_created,
            imports_updated=imports_updated,
            lines_created=lines_created,
            lines_updated=lines_updated,
            periods=sorted(periods),
            sheets_imported=sheets_imported,
        )

    @transaction.atomic
    def import_workbook(self, workbook_path: str | Path) -> TrustedBudgetDetailImportSummary:
        path = Path(workbook_path).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(path)

        expected_sheets = self.expected_sheets_for_file(path.name)
        workbook_formula = load_workbook(path, read_only=True, data_only=False)
        workbook_values = load_workbook(path, read_only=True, data_only=True)
        missing_sheets = [sheet_name for sheet_name in expected_sheets if sheet_name not in workbook_formula.sheetnames]
        if missing_sheets:
            raise ValueError(
                f"{path.name} no contiene todas las hojas esperadas: {', '.join(missing_sheets)}."
            )

        imports_created = 0
        imports_updated = 0
        lines_created = 0
        lines_updated = 0
        periods: set[str] = set()
        sheets_imported: list[str] = []

        for sheet_name, kind in expected_sheets.items():
            result = self._import_sheet(
                path=path,
                workbook_formula=workbook_formula,
                workbook_values=workbook_values,
                sheet_name=sheet_name,
                kind=kind,
            )
            imports_created += result[0]
            imports_updated += result[1]
            lines_created += result[2]
            lines_updated += result[3]
            periods.update(result[4])
            sheets_imported.append(f"{path.name}::{sheet_name}")

        return TrustedBudgetDetailImportSummary(
            imports_created=imports_created,
            imports_updated=imports_updated,
            lines_created=lines_created,
            lines_updated=lines_updated,
            periods=sorted(periods),
            sheets_imported=sheets_imported,
        )


class BudgetGeneralAuditService:
    SALES_BRANCH_SHEETS = ["GUAMUCHIL", "EL TUNEL", "PLAZA NIO", "CRUCERO", "COLOSIO", "GLORIAS", "PAYAN", "LEYVA", "MATRIZ"]
    NOMINA_AREA_SHEETS = ["VENTAS", "PRODUCCCION", "LOGISTICA", "ADMINISTRACION"]
    ADMIN_NOMINA_EXPECTED = {
        "SUELDO": "SUELDO",
        "DIAS FESTIVOS": "FESTIVO",
        "VACACIONES": "VACACIONES",
        "PRIMA VACACIONAL": "PRIMA VACACIONES",
        "BONO POR RESULTADOS (DESPENSA)": "BONOS POR RESULTADOS",
        "BONO PUNTUALIDAD": "BONOS PUNTUALIDAD",
        "BONO POR ASISTENCIA": "BONOS ASISTENCIA",
        "IMSS": "IMSS",
        "INFONAVIT-RCV": "INFONAVIT",
        "AGUINALDO": "AGUINALDO",
        "UTILIDADES": "UTILIDADES",
        "PLAYERA": "PLAYERAS",
        "MANDIL": "MANDIL",
        "CAMISA MUJER": "CAMISA MUJER",
        "CAMISA HOMBRE": "CAMISA HOMBRE",
        "GORRA": "GORRA",
    }
    NOMINA_AREA_CONCEPT_ALIASES = {
        "SUELDO": "SUELDO",
        "DIAS FESTIVOS": "FESTIVO",
        "VACACIONES": "VACACIONES",
        "PRIMA VACACIONAL": "PRIMA VACACIONES",
        "PRIMA VACACIONES": "PRIMA VACACIONES",
        "BONO POR RESULTADOS (DESPENSA)": "BONOS POR RESULTADOS",
        "BONO PUNTUALIDAD": "BONOS PUNTUALIDAD",
        "BONO POR ASISTENCIA": "BONOS ASISTENCIA",
        "IMSS": "IMSS",
        "INFONAVIT-RCV": "INFONAVIT",
        "AGUINALDO": "AGUINALDO",
        "UTILIDADES": "UTILIDADES",
        "PLAYERA": "PLAYERAS",
        "MANDIL": "MANDIL",
        "CAMISA MUJER": "CAMISA MUJER",
        "CAMISA HOMBRE": "CAMISA HOMBRE",
        "GORRA": "GORRA",
    }
    EXTERNAL_REF_PATTERN = re.compile(r"\[(\d+)\]'?([^']+?)'?!([A-Z]+\d+)")

    def __init__(self) -> None:
        self._general = GeneralBudgetImportService()

    def _normalize(self, text: str) -> str:
        return " ".join((text or "").strip().upper().replace(".", "").split())

    def _parse_sheet_monthly_budgets(self, worksheet) -> dict[str, dict[str, object]]:
        _, monthly_templates, data_start_row = self._general._parse_layout(worksheet)
        rows: dict[str, dict[str, object]] = {}
        for row_index in range(data_start_row, worksheet.max_row + 1):
            raw_account = worksheet.cell(row_index, 1).value
            raw_concept = worksheet.cell(row_index, 2).value
            account_code = str(raw_account or "").strip()
            concept = str(raw_concept or "").strip()
            if not isinstance(raw_concept, str) and account_code:
                concept = account_code
            elif not concept and account_code:
                concept = account_code
            normalized = self._normalize(concept)
            if not normalized:
                continue
            payload = rows.setdefault(
                normalized,
                {
                    "concept": concept,
                    "row_index": row_index,
                    "months": {month_number: 0.0 for month_number, _ in monthly_templates},
                },
            )
            for month_number, columns in monthly_templates:
                budget_col = columns.get("budget")
                if budget_col:
                    payload["months"][month_number] += float(_to_decimal(worksheet.cell(row_index, budget_col).value))
        return rows

    def _month_label(self, month_number: int) -> str:
        return date(2000, month_number, 1).strftime("%b").lower()

    def _audit_general_vs_detail_all_months(
        self,
        *,
        general_ws,
        detail_workbook,
        detail_sheets: list[str],
        detail_alias_map: dict[str, str] | None = None,
        ignore_zero_only: bool = True,
    ) -> dict[str, object]:
        detail_alias_map = detail_alias_map or {}
        general_rows = self._parse_sheet_monthly_budgets(general_ws)
        detail_rows_by_sheet = {
            sheet_name: self._parse_sheet_monthly_budgets(detail_workbook[sheet_name])
            for sheet_name in detail_sheets
            if sheet_name in detail_workbook.sheetnames
        }
        concepts = sorted(general_rows.keys())
        ok_count = 0
        mismatch_count = 0
        missing_count = 0
        rows: list[dict[str, object]] = []
        mismatches: list[dict[str, object]] = []

        for general_key in concepts:
            general_payload = general_rows[general_key]
            detail_key = self._normalize(detail_alias_map.get(general_key, general_key))
            detail_months = {month: 0.0 for month in general_payload["months"].keys()}
            found_in_detail = False
            for sheet_rows in detail_rows_by_sheet.values():
                detail_payload = sheet_rows.get(detail_key)
                if not detail_payload:
                    continue
                found_in_detail = True
                for month, value in detail_payload["months"].items():
                    detail_months[month] = detail_months.get(month, 0.0) + float(value or 0)

            if ignore_zero_only and not found_in_detail and all(abs(v) < 0.0001 for v in general_payload["months"].values()):
                continue

            month_diffs = []
            max_abs_diff = 0.0
            for month, general_value in general_payload["months"].items():
                detail_value = detail_months.get(month, 0.0)
                diff = round(float(general_value or 0) - float(detail_value or 0), 2)
                max_abs_diff = max(max_abs_diff, abs(diff))
                if abs(diff) > 0.01:
                    month_diffs.append(
                        {
                            "month": month,
                            "month_label": self._month_label(month),
                            "general": round(float(general_value or 0), 2),
                            "detail": round(float(detail_value or 0), 2),
                            "diff": diff,
                        }
                    )

            if not found_in_detail:
                status = "missing_in_detail"
                missing_count += 1
            elif month_diffs:
                status = "mismatch"
                mismatch_count += 1
            else:
                status = "ok"
                ok_count += 1

            row = {
                "concept": general_payload["concept"],
                "status": status,
                "row_index": general_payload["row_index"],
                "max_abs_diff": round(max_abs_diff, 2),
                "month_diffs": month_diffs,
            }
            rows.append(row)
            if status != "ok":
                mismatches.append(row)

        return {
            "ok_count": ok_count,
            "mismatch_count": mismatch_count,
            "missing_count": missing_count,
            "reviewed_concepts": len(rows),
            "mismatches": mismatches,
            "rows": rows,
        }

    def _resolve_external_targets(self, workbook) -> dict[str, Path]:
        mapping: dict[str, Path] = {}
        for index, link in enumerate(getattr(workbook, "_external_links", []), start=1):
            target = getattr(getattr(link, "file_link", None), "Target", "") or ""
            target_name = Path(unquote(target)).name
            if target_name:
                mapping[str(index)] = Path(target_name)
        return mapping

    def _audit_admin_general_external_links(self, folder: Path) -> dict[str, object]:
        admin_path = folder / "PRESUPUESTO ADMINISTRACIÓN 2026.xlsx"
        if not admin_path.exists():
            return {"reviewed_cells": 0, "mismatch_count": 0, "missing_source_count": 0, "mismatches": []}

        admin_formula_wb = load_workbook(admin_path, data_only=False, read_only=True)
        admin_values_wb = load_workbook(admin_path, data_only=True, read_only=True)
        admin_formula_ws = admin_formula_wb["GENERAL"]
        admin_values_ws = admin_values_wb["GENERAL"]
        _, monthly_templates, data_start_row = self._general._parse_layout(admin_formula_ws)
        external_targets = self._resolve_external_targets(admin_formula_wb)
        source_books: dict[str, object] = {}
        reviewed_cells = 0
        mismatch_count = 0
        missing_source_count = 0
        mismatches: list[dict[str, object]] = []

        for row_index in range(data_start_row, admin_formula_ws.max_row + 1):
            concept = str(admin_values_ws.cell(row_index, 2).value or "").strip() or str(admin_values_ws.cell(row_index, 1).value or "").strip()
            if not concept:
                continue
            for month_number, columns in monthly_templates:
                budget_col = columns.get("budget")
                if not budget_col:
                    continue
                formula = admin_formula_ws.cell(row_index, budget_col).value
                if not isinstance(formula, str) or "[" not in formula or "!" not in formula:
                    continue
                match = self.EXTERNAL_REF_PATTERN.search(formula)
                if not match:
                    continue
                reviewed_cells += 1
                link_index, sheet_name, cell_ref = match.groups()
                source_name = external_targets.get(link_index)
                if not source_name:
                    missing_source_count += 1
                    mismatches.append(
                        {
                            "concept": concept,
                            "month": month_number,
                            "month_label": self._month_label(month_number),
                            "status": "missing_external_link",
                            "formula": formula,
                        }
                    )
                    continue
                source_path = folder / source_name
                if not source_path.exists():
                    missing_source_count += 1
                    mismatches.append(
                        {
                            "concept": concept,
                            "month": month_number,
                            "month_label": self._month_label(month_number),
                            "status": "missing_source_file",
                            "source_file": source_name.name,
                            "formula": formula,
                        }
                    )
                    continue
                source_book = source_books.get(source_name.name)
                if source_book is None:
                    source_book = load_workbook(source_path, data_only=True, read_only=True)
                    source_books[source_name.name] = source_book
                if sheet_name not in source_book.sheetnames:
                    mismatch_count += 1
                    mismatches.append(
                        {
                            "concept": concept,
                            "month": month_number,
                            "month_label": self._month_label(month_number),
                            "status": "missing_source_sheet",
                            "source_file": source_name.name,
                            "source_sheet": sheet_name,
                            "formula": formula,
                        }
                    )
                    continue
                source_value = float(source_book[sheet_name][cell_ref].value or 0)
                current_value = float(admin_values_ws.cell(row_index, budget_col).value or 0)
                if abs(round(current_value - source_value, 2)) > 0.01:
                    mismatch_count += 1
                    mismatches.append(
                        {
                            "concept": concept,
                            "month": month_number,
                            "month_label": self._month_label(month_number),
                            "status": "value_mismatch",
                            "source_file": source_name.name,
                            "source_sheet": sheet_name,
                            "source_cell": cell_ref,
                            "formula": formula,
                            "general": round(current_value, 2),
                            "source": round(source_value, 2),
                            "diff": round(current_value - source_value, 2),
                        }
                    )
        return {
            "reviewed_cells": reviewed_cells,
            "mismatch_count": mismatch_count,
            "missing_source_count": missing_source_count,
            "mismatches": mismatches,
        }

    def _audit_admin_general_nomina_links(self, folder: Path) -> list[dict[str, object]]:
        admin_path = folder / "PRESUPUESTO ADMINISTRACIÓN 2026.xlsx"
        nomina_path = folder / "PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx"
        if not admin_path.exists() or not nomina_path.exists():
            return []

        admin_formula = load_workbook(admin_path, data_only=False, read_only=True)["GENERAL"]
        admin_values = load_workbook(admin_path, data_only=True, read_only=True)["GENERAL"]
        nomina_values = load_workbook(nomina_path, data_only=True, read_only=True)["GENERAL"]

        findings: list[dict[str, object]] = []
        for row_index in range(5, admin_values.max_row + 1):
            concept = str(admin_values.cell(row_index, 2).value or "").strip()
            if not concept:
                continue
            formula = str(admin_formula.cell(row_index, 6).value or "")
            if "[4]GENERAL!B" not in formula:
                continue
            match = re.search(r"\[4\]GENERAL!B(\d+)", formula)
            if not match:
                continue
            target_row = int(match.group(1))
            target_label = str(nomina_values.cell(target_row, 1).value or "").strip()
            expected_label = self.ADMIN_NOMINA_EXPECTED.get(self._normalize(concept), self._normalize(concept))
            target_normalized = self._normalize(target_label)
            status = "ok" if target_normalized == expected_label else "reference_mismatch"
            findings.append(
                {
                    "row_index": row_index,
                    "concept": concept,
                    "formula": formula,
                    "target_row": target_row,
                    "target_label": target_label,
                    "target_budget_enero": float(nomina_values.cell(target_row, 2).value or 0),
                    "expected_label": expected_label,
                    "status": status,
                }
            )
        return findings

    def audit_folder(self, folder_path: str | Path) -> dict[str, object]:
        folder = Path(folder_path).expanduser().resolve()
        sales_wb = load_workbook(folder / "PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx", data_only=True, read_only=True)
        nomina_wb = load_workbook(folder / "PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx", data_only=True, read_only=True)

        def d(v) -> float:
            return float(v or 0)

        sales_checks = []
        for concept in ["Sueldo", "Arrendamiento local"]:
            general_row = None
            general_ws = sales_wb["GENERAL"]
            for r in range(1, 240):
                if str(general_ws.cell(r, 2).value or "").strip() == concept:
                    general_row = r
                    break
            detail_enero = 0.0
            detail_febrero = 0.0
            for sheet_name in self.SALES_BRANCH_SHEETS:
                if sheet_name not in sales_wb.sheetnames:
                    continue
                ws = sales_wb[sheet_name]
                for r in range(1, 240):
                    if str(ws.cell(r, 2).value or "").strip() == concept:
                        detail_enero += d(ws.cell(r, 6).value)
                        detail_febrero += d(ws.cell(r, 9).value)
                        break
            sales_checks.append(
                {
                    "concept": concept,
                    "general_enero": d(general_ws.cell(general_row, 6).value) if general_row else None,
                    "detail_enero": detail_enero,
                    "general_febrero": d(general_ws.cell(general_row, 9).value) if general_row else None,
                    "detail_febrero": detail_febrero,
                }
            )

        nomina_checks = []
        general_ws = nomina_wb["GENERAL"]
        general_month_cols = {
            "enero": {"budget": 2},
            "febrero": {"budget": 5},
        }
        concept_map = {
            "SUELDO": {"GENERAL": "SUELDO", "areas": "SUELDO"},
            "VACACIONES": {"GENERAL": "VACACIONES", "areas": "VACACIONES"},
            "PRIMA VACACIONAL": {"GENERAL": "PRIMA VACACIONES", "areas": "PRIMA VACACIONES"},
        }
        for label, mapping in concept_map.items():
            general_row = None
            for r in range(1, 120):
                if self._normalize(str(general_ws.cell(r, 1).value or "")) == self._normalize(mapping["GENERAL"]):
                    general_row = r
                    break
            payload = {"concept": label}
            for month_label, month_cols in [("enero", general_month_cols["enero"]), ("febrero", general_month_cols["febrero"])]:
                general_value = d(general_ws.cell(general_row, month_cols["budget"]).value) if general_row else 0.0
                detail_value = 0.0
                for sheet_name in self.NOMINA_AREA_SHEETS:
                    if sheet_name not in nomina_wb.sheetnames:
                        continue
                    ws = nomina_wb[sheet_name]
                    concept_col = 2 if sheet_name == "VENTAS" else 1
                    budget_col = 3 if sheet_name == "VENTAS" else 2
                    if month_label == "febrero":
                        budget_col = 6 if sheet_name == "VENTAS" else 5
                    for r in range(1, 120):
                        if self._normalize(str(ws.cell(r, concept_col).value or "")) == self._normalize(mapping["areas"]):
                            detail_value += d(ws.cell(r, budget_col).value)
                            break
                payload[f"general_{month_label}"] = general_value
                payload[f"detail_{month_label}"] = detail_value
            nomina_checks.append(payload)

        sales_full_audit = self._audit_general_vs_detail_all_months(
            general_ws=sales_wb["GENERAL"],
            detail_workbook=sales_wb,
            detail_sheets=self.SALES_BRANCH_SHEETS,
        )
        nomina_full_audit = self._audit_general_vs_detail_all_months(
            general_ws=nomina_wb["GENERAL"],
            detail_workbook=nomina_wb,
            detail_sheets=self.NOMINA_AREA_SHEETS,
            detail_alias_map=self.NOMINA_AREA_CONCEPT_ALIASES,
        )

        return {
            "sales_general_vs_branches": sales_checks,
            "sales_general_vs_branches_full": sales_full_audit,
            "nomina_general_vs_areas": nomina_checks,
            "nomina_general_vs_areas_full": nomina_full_audit,
            "admin_general_nomina_links": self._audit_admin_general_nomina_links(folder),
            "admin_general_external_links_full": self._audit_admin_general_external_links(folder),
        }


@dataclass
class BudgetAuditMaterializationSummary:
    total_lines: int
    ok_lines: int
    deviation_lines: int
    bad_formula_lines: int
    missing_detail_lines: int
    excluded_total_lines: int
    excluded_extra_lines: int
    excluded_duplicate_lines: int


class BudgetAuditMaterializationService:
    EXCLUDED_TOTAL_CONCEPTS = {
        "INGRESOS",
        "EGRESOS",
        "COSTOS",
        "UTILIDAD BRUTA",
        "UTILIDAD O PERDIDA",
        "UTILIDAD O PÉRDIDA",
        "VENTA COMPLEMENTOS",
        "VENTA POSTRES",
        "PRODUCCIÓN",
        "LOGÍSTICA",
        "TOTAL GASTOS",
        "TOTAL GASTOS VENTAS",
        "TOTAL POR MES",
        "TOTALES",
        "UNIFORMES",
    }
    EXCLUDED_EXTRA_CONCEPTS = {
        "APERTURA SUCURSAL",
        "ADQUISICIÓN DE EQUIPO/MAQUINARIA",
        "ADQUISICION DE EQUIPO/MAQUINARIA",
    }
    DUPLICATE_PAYROLL_CONCEPTS = {
        "SUELDO",
        "FESTIVO",
        "VACACIONES",
        "PRIMA VACACIONAL",
        "PRIMA VACACIONES",
        "BONO POR RESULTADOS (DESPENSA)",
        "BONOS POR RESULTADOS",
        "BONO PUNTUALIDAD",
        "BONOS PUNTUALIDAD",
        "BONO POR ASISTENCIA",
        "BONOS ASISTENCIA",
        "IMSS",
        "INFONAVIT",
        "INFONAVIT-RCV",
        "AGUINALDO",
        "UTILIDADES",
        "PLAYERA",
        "PLAYERAS",
        "MANDIL",
        "POLO",
        "CAMISA MUJER",
        "CAMISA HOMBRE",
        "GORRA",
    }

    def __init__(self) -> None:
        self._audit = BudgetGeneralAuditService()

    def _normalized_concept(self, value: str) -> str:
        return self._audit._normalize(value)

    def _detail_status(self, line: PresupuestoLineaMensual) -> tuple[str, str]:
        concept = self._normalized_concept(line.concept)
        kind = str((line.metadata or {}).get("kind") or (line.importacion.metadata or {}).get("kind") or "").strip()
        if not concept:
            return PresupuestoLineaMensual.AUDIT_MISSING_DETAIL, "detalle"
        if concept in self.EXCLUDED_TOTAL_CONCEPTS or concept.startswith("TOTAL "):
            return PresupuestoLineaMensual.AUDIT_EXCLUDED_TOTAL, "detail_rules"
        if concept in self.EXCLUDED_EXTRA_CONCEPTS:
            return PresupuestoLineaMensual.AUDIT_EXCLUDED_EXTRA, "detail_rules"
        if kind in {"production_budget", "logistics_budget"} and concept in self.DUPLICATE_PAYROLL_CONCEPTS:
            return PresupuestoLineaMensual.AUDIT_EXCLUDED_DUPLICATE, "detail_rules"
        return PresupuestoLineaMensual.AUDIT_OK, "detail_rules"

    def _apply_general_audit(
        self,
        *,
        fuente_nombre: str,
        mismatches: list[dict[str, object]],
        ok_rows: list[dict[str, object]],
        source: str,
    ) -> None:
        for row in ok_rows:
            for month in range(1, 13):
                PresupuestoLineaMensual.objects.filter(
                    importacion__fuente_nombre=fuente_nombre,
                    importacion__sheet_name="GENERAL",
                    period__month=month,
                    concept=row["concept"],
                ).update(audit_status=PresupuestoLineaMensual.AUDIT_OK, audit_source=source)
        for row in mismatches:
            status = row["status"]
            if status == "missing_in_detail":
                audit_status = PresupuestoLineaMensual.AUDIT_MISSING_DETAIL
            else:
                audit_status = PresupuestoLineaMensual.AUDIT_DEVIATION
            month_set = {month_diff["month"] for month_diff in row.get("month_diffs", [])} or set(range(1, 13))
            PresupuestoLineaMensual.objects.filter(
                importacion__fuente_nombre=fuente_nombre,
                importacion__sheet_name="GENERAL",
                period__month__in=month_set,
                concept=row["concept"],
            ).update(audit_status=audit_status, audit_source=source)

    @transaction.atomic
    def materialize(self, folder_path: str | Path) -> BudgetAuditMaterializationSummary:
        folder = Path(folder_path).expanduser().resolve()
        payload = self._audit.audit_folder(folder)

        PresupuestoLineaMensual.objects.filter(importacion__tipo=PresupuestoImport.TIPO_GENERAL).update(
            audit_status=PresupuestoLineaMensual.AUDIT_PENDING,
            audit_source="",
        )

        for line in PresupuestoLineaMensual.objects.filter(importacion__tipo=PresupuestoImport.TIPO_DETALLE).select_related("importacion"):
            status, source = self._detail_status(line)
            if line.audit_status != status or line.audit_source != source:
                PresupuestoLineaMensual.objects.filter(pk=line.pk).update(audit_status=status, audit_source=source)

        sales_rows = payload["sales_general_vs_branches_full"]["rows"]
        self._apply_general_audit(
            fuente_nombre="PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx",
            mismatches=[row for row in sales_rows if row["status"] != "ok"],
            ok_rows=[row for row in sales_rows if row["status"] == "ok"],
            source="general_vs_branch_detail",
        )

        nomina_rows = payload["nomina_general_vs_areas_full"]["rows"]
        self._apply_general_audit(
            fuente_nombre="PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx",
            mismatches=[row for row in nomina_rows if row["status"] != "ok"],
            ok_rows=[row for row in nomina_rows if row["status"] == "ok"],
            source="general_vs_area_detail",
        )

        admin_general_qs = PresupuestoLineaMensual.objects.filter(
            importacion__fuente_nombre="PRESUPUESTO ADMINISTRACIÓN 2026.xlsx",
            importacion__sheet_name="GENERAL",
        )
        admin_general_qs.update(audit_status=PresupuestoLineaMensual.AUDIT_PENDING, audit_source="admin_general_control")

        for row in payload["admin_general_external_links_full"]["mismatches"]:
            target_qs = admin_general_qs.filter(
                period__month=row["month"],
                concept=row["concept"],
            )
            status = row["status"]
            if status in {"missing_external_link", "missing_source_file", "missing_source_sheet"}:
                audit_status = PresupuestoLineaMensual.AUDIT_MISSING_DETAIL
            else:
                audit_status = PresupuestoLineaMensual.AUDIT_DEVIATION
            target_qs.update(audit_status=audit_status, audit_source="admin_general_external_links")

        bad_formula_concepts = {
            row["concept"] for row in payload["admin_general_nomina_links"] if row["status"] != "ok"
        }
        if bad_formula_concepts:
            admin_general_qs.filter(concept__in=bad_formula_concepts).update(
                audit_status=PresupuestoLineaMensual.AUDIT_BAD_FORMULA,
                audit_source="admin_general_nomina_links",
            )

        counts = PresupuestoLineaMensual.objects.values("audit_status").annotate(total=models.Count("id"))
        counter = {row["audit_status"]: int(row["total"] or 0) for row in counts}
        return BudgetAuditMaterializationSummary(
            total_lines=sum(counter.values()),
            ok_lines=counter.get(PresupuestoLineaMensual.AUDIT_OK, 0),
            deviation_lines=counter.get(PresupuestoLineaMensual.AUDIT_DEVIATION, 0),
            bad_formula_lines=counter.get(PresupuestoLineaMensual.AUDIT_BAD_FORMULA, 0),
            missing_detail_lines=counter.get(PresupuestoLineaMensual.AUDIT_MISSING_DETAIL, 0),
            excluded_total_lines=counter.get(PresupuestoLineaMensual.AUDIT_EXCLUDED_TOTAL, 0),
            excluded_extra_lines=counter.get(PresupuestoLineaMensual.AUDIT_EXCLUDED_EXTRA, 0),
            excluded_duplicate_lines=counter.get(PresupuestoLineaMensual.AUDIT_EXCLUDED_DUPLICATE, 0),
        )
