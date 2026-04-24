from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from core.audit import log_event
from core.models import Sucursal
from reportes.models import CategoriaGasto, CentroCosto, GastoOperativoMensual, ProyectoInversion
from reportes.services_investment_projects import ProyectoInversionRefreshService
from reportes.services_operating_finance import OperatingFinanceBootstrapService


SUPPORTED_TYPE_VALUES = {
    "REAL": GastoOperativoMensual.TIPO_DATO_REAL,
    "PRESUPUESTO": GastoOperativoMensual.TIPO_DATO_PRESUPUESTO,
}
DEFAULT_SHEET_NAME = "GastosSucursal"
TOTAL_BRANCH_CATEGORY_CODE = "OPEX_TOTAL_SUC"
NON_REAL_POLICY_IGNORE = "ignore"
NON_REAL_POLICY_REJECT = "reject"


def _normalize_text(value) -> str:
    return str(value or "").strip()


def _normalize_key(value) -> str:
    return "".join(ch for ch in _normalize_text(value).upper() if ch.isalnum())


def _parse_period(raw_value) -> date:
    value = _normalize_text(raw_value)
    if not value:
        raise ValueError("periodo vacío")
    try:
        period = date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError(f"periodo inválido '{value}', se espera YYYY-MM-01") from exc
    if period.day != 1:
        raise ValueError(f"periodo inválido '{value}', el día debe ser 01")
    return period


def _parse_amount(raw_value) -> Decimal:
    value = _normalize_text(raw_value)
    if not value:
        raise ValueError("monto vacío")
    normalized = value.replace(",", "")
    try:
        return Decimal(normalized)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"monto inválido '{value}'") from exc


def _detect_row_type(*, raw_value, workbook_name: str, sheet_name: str) -> str:
    explicit = _normalize_key(raw_value)
    if explicit in SUPPORTED_TYPE_VALUES:
        return SUPPORTED_TYPE_VALUES[explicit]

    combined = _normalize_key(f"{workbook_name} {sheet_name}")
    if "PRESUPUESTO" in combined:
        return GastoOperativoMensual.TIPO_DATO_PRESUPUESTO
    if any(token in combined for token in {"REAL", "REALES", "ACTUAL", "ACTUALES"}):
        return GastoOperativoMensual.TIPO_DATO_REAL
    raise ValueError("tipo_dato no identificable; agrega columna tipo_dato=REAL/PRESUPUESTO o renombra el archivo")


@dataclass
class BranchRealOperatingExpenseImportError:
    row_number: int
    field: str
    message: str


@dataclass
class BranchRealOperatingExpenseImportSummary:
    processed_rows: int = 0
    created: int = 0
    updated: int = 0
    skipped_non_real: int = 0
    project_refresh_count: int = 0
    project_ids: list[int] = field(default_factory=list)
    periods: list[str] = field(default_factory=list)
    affected_branches: list[str] = field(default_factory=list)
    errors: list[BranchRealOperatingExpenseImportError] = field(default_factory=list)

    @property
    def loaded_rows(self) -> int:
        return self.created + self.updated


@dataclass
class _NormalizedExpenseRow:
    row_number: int
    sucursal: Sucursal
    period: date
    amount: Decimal
    tipo_dato: str
    category: CategoriaGasto
    external_key: str
    source_file: str
    source_label: str
    comment: str


class BranchRealOperatingExpenseImportValidationError(ValueError):
    def __init__(self, summary: BranchRealOperatingExpenseImportSummary):
        self.summary = summary
        message = "; ".join(
            f"fila {item.row_number} [{item.field}]: {item.message}"
            for item in summary.errors
        ) or "Validación de carga falló."
        super().__init__(message)


class BranchRealOperatingExpenseImportService:
    REQUIRED_HEADERS = {"sucursal", "periodo", "monto"}

    def __init__(self) -> None:
        self._refresh_service = ProyectoInversionRefreshService()

    def _branch_index(self) -> dict[str, Sucursal]:
        index: dict[str, Sucursal] = {}
        for branch in Sucursal.objects.order_by("codigo", "nombre"):
            index[_normalize_key(branch.codigo)] = branch
            index[_normalize_key(branch.nombre)] = branch
        return index

    def _cost_center_index(self) -> dict[int, CentroCosto]:
        centers = (
            CentroCosto.objects.filter(tipo=CentroCosto.TIPO_SUCURSAL, sucursal__isnull=False)
            .select_related("sucursal")
            .order_by("codigo")
        )
        return {center.sucursal_id: center for center in centers}

    def _category_index(self) -> dict[str, CategoriaGasto]:
        return {
            category.codigo: category
            for category in CategoriaGasto.objects.filter(capa_objetivo=CategoriaGasto.CAPA_SUCURSAL)
        }

    def _load_sheet_rows(self, workbook_path: Path, *, sheet_name: str | None = None) -> tuple[str, list[dict[str, object]]]:
        workbook = load_workbook(filename=workbook_path, data_only=True)
        selected_sheet = sheet_name or (DEFAULT_SHEET_NAME if DEFAULT_SHEET_NAME in workbook.sheetnames else workbook.sheetnames[0])
        if selected_sheet not in workbook.sheetnames:
            raise ValueError(f"La hoja '{selected_sheet}' no existe en el archivo.")
        ws = workbook[selected_sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"La hoja '{selected_sheet}' está vacía.")
        headers = [_normalize_text(value) for value in rows[0]]
        missing = self.REQUIRED_HEADERS - set(headers)
        if missing:
            raise ValueError(f"Faltan columnas requeridas: {', '.join(sorted(missing))}.")
        header_map = {header: index for index, header in enumerate(headers)}
        payload_rows: list[dict[str, object]] = []
        for row_number, values in enumerate(rows[1:], start=2):
            if not any(values):
                continue
            payload_rows.append(
                {
                    "row_number": row_number,
                    "sucursal": values[header_map["sucursal"]],
                    "periodo": values[header_map["periodo"]],
                    "monto": values[header_map["monto"]],
                    "tipo_dato": values[header_map["tipo_dato"]] if "tipo_dato" in header_map else "",
                    "categoria_gasto": values[header_map["categoria_gasto"]] if "categoria_gasto" in header_map else "",
                    "external_key": values[header_map["external_key"]] if "external_key" in header_map else "",
                    "comentario": values[header_map["comentario"]] if "comentario" in header_map else "",
                }
            )
        return selected_sheet, payload_rows

    def _validate_rows(
        self,
        *,
        rows: list[dict[str, object]],
        workbook_name: str,
        sheet_name: str,
        target_year: int,
        non_real_policy: str,
        allow_open_month_real: bool,
    ) -> tuple[BranchRealOperatingExpenseImportSummary, list[_NormalizedExpenseRow]]:
        summary = BranchRealOperatingExpenseImportSummary()
        branch_index = self._branch_index()
        center_index = self._cost_center_index()
        category_index = self._category_index()
        default_category = category_index.get(TOTAL_BRANCH_CATEGORY_CODE)
        if default_category is None:
            raise ValueError(
                f"No existe la categoría {TOTAL_BRANCH_CATEGORY_CODE}; ejecuta bootstrap_operating_finance primero."
            )

        normalized_rows: list[_NormalizedExpenseRow] = []
        seen_keys: set[tuple[str, date, str, str]] = set()
        seen_external_keys: set[str] = set()
        branch_period_categories: dict[tuple[str, date], set[str]] = {}
        for payload in rows:
            row_number = int(payload["row_number"])
            summary.processed_rows += 1
            try:
                branch_key = _normalize_key(payload["sucursal"])
                branch = branch_index.get(branch_key)
                if branch is None:
                    raise ValueError(f"sucursal desconocida '{_normalize_text(payload['sucursal'])}'")
                if branch.id not in center_index:
                    raise ValueError(f"la sucursal '{branch.codigo}' no tiene centro de costo configurado")

                period = _parse_period(payload["periodo"])
                if period.year != target_year:
                    raise ValueError(f"periodo fuera de alcance: {period.isoformat()} (solo {target_year})")

                amount = _parse_amount(payload["monto"])
                row_type = _detect_row_type(
                    raw_value=payload.get("tipo_dato"),
                    workbook_name=workbook_name,
                    sheet_name=sheet_name,
                )
                current_period = timezone.localdate().replace(day=1)
                if (
                    row_type == GastoOperativoMensual.TIPO_DATO_REAL
                    and period == current_period
                    and not allow_open_month_real
                ):
                    raise ValueError(
                        "no se permite cargar gasto REAL del mes en curso por canal automático; "
                        "usa captura manual o importación web explícita"
                    )
                if row_type != GastoOperativoMensual.TIPO_DATO_REAL:
                    if non_real_policy == NON_REAL_POLICY_IGNORE:
                        summary.skipped_non_real += 1
                        continue
                    raise ValueError("tipo_dato PRESUPUESTO no permitido para esta carga")

                category_code = _normalize_text(payload.get("categoria_gasto")) or TOTAL_BRANCH_CATEGORY_CODE
                category = category_index.get(category_code)
                if category is None:
                    raise ValueError(f"categoria_gasto desconocida '{category_code}'")

                duplicate_key = (branch.codigo, period, category.codigo, row_type)
                if duplicate_key in seen_keys:
                    raise ValueError("duplicado en archivo para sucursal/periodo/categoria/tipo_dato")
                seen_keys.add(duplicate_key)

                external_key = _normalize_text(payload.get("external_key")) or (
                    f"BRANCH_REAL_OPEX|{branch.codigo}|{period.isoformat()}|{category.codigo}|{row_type}"
                )
                if external_key in seen_external_keys:
                    raise ValueError(f"external_key duplicado '{external_key}'")
                seen_external_keys.add(external_key)
                branch_period_categories.setdefault((branch.codigo, period), set()).add(category.codigo)
                normalized_rows.append(
                    _NormalizedExpenseRow(
                        row_number=row_number,
                        sucursal=branch,
                        period=period,
                        amount=amount,
                        tipo_dato=row_type,
                        category=category,
                        external_key=external_key,
                        source_file=workbook_name,
                        source_label=f"{workbook_name}:{sheet_name}",
                        comment=_normalize_text(payload.get("comentario"))
                        or f"Carga real sucursal {branch.codigo} {period:%Y-%m}",
                    )
                )
            except ValueError as exc:
                field = "row"
                if "sucursal" in str(exc):
                    field = "sucursal"
                elif "periodo" in str(exc):
                    field = "periodo"
                elif "monto" in str(exc):
                    field = "monto"
                elif "tipo_dato" in str(exc):
                    field = "tipo_dato"
                elif "categoria_gasto" in str(exc):
                    field = "categoria_gasto"
                summary.errors.append(
                    BranchRealOperatingExpenseImportError(
                        row_number=row_number,
                        field=field,
                        message=str(exc),
                    )
                )
        for (branch_code, period), categories in sorted(branch_period_categories.items()):
            if TOTAL_BRANCH_CATEGORY_CODE in categories and len(categories) > 1:
                summary.errors.append(
                    BranchRealOperatingExpenseImportError(
                        row_number=0,
                        field="categoria_gasto",
                        message=(
                            f"mezcla inválida para {branch_code} {period.isoformat()}: "
                            "no combines OPEX_TOTAL_SUC con detalle categorizado"
                        ),
                    )
                )
        if summary.errors:
            raise BranchRealOperatingExpenseImportValidationError(summary)
        return summary, normalized_rows

    @transaction.atomic
    def import_workbook(
        self,
        workbook_path: str | Path,
        *,
        target_year: int = 2026,
        non_real_policy: str = NON_REAL_POLICY_IGNORE,
        refresh_projects: bool = True,
        refresh_until: date | None = None,
        user=None,
        sheet_name: str | None = None,
        allow_open_month_real: bool = True,
    ) -> BranchRealOperatingExpenseImportSummary:
        if non_real_policy not in {NON_REAL_POLICY_IGNORE, NON_REAL_POLICY_REJECT}:
            raise ValueError(f"non_real_policy inválida: {non_real_policy}")
        OperatingFinanceBootstrapService().bootstrap()
        workbook = Path(workbook_path).expanduser().resolve()
        if not workbook.exists():
            raise FileNotFoundError(workbook)

        selected_sheet, payload_rows = self._load_sheet_rows(workbook, sheet_name=sheet_name)
        summary, normalized_rows = self._validate_rows(
            rows=payload_rows,
            workbook_name=workbook.name,
            sheet_name=selected_sheet,
            target_year=target_year,
            non_real_policy=non_real_policy,
            allow_open_month_real=allow_open_month_real,
        )
        center_index = self._cost_center_index()
        affected_branch_codes: set[str] = set()
        affected_periods: set[str] = set()

        for row in normalized_rows:
            center = center_index[row.sucursal.id]
            _, was_created = GastoOperativoMensual.objects.update_or_create(
                external_key=row.external_key,
                defaults={
                    "periodo": row.period,
                    "centro_costo": center,
                    "categoria_gasto": row.category,
                    "monto": row.amount,
                    "tipo_dato": row.tipo_dato,
                    "fuente": GastoOperativoMensual.FUENTE_IMPORTADA,
                    "es_estimado": False,
                    "comentario": row.comment,
                    "archivo_soporte": row.source_file,
                    "capturado_por": user if getattr(user, "is_authenticated", False) else None,
                },
            )
            summary.created += int(was_created)
            summary.updated += int(not was_created)
            affected_branch_codes.add(row.sucursal.codigo)
            affected_periods.add(row.period.isoformat())

        summary.affected_branches = sorted(affected_branch_codes)
        summary.periods = sorted(affected_periods)

        if refresh_projects and affected_branch_codes:
            projects = list(
                ProyectoInversion.objects.filter(sucursal_relacionada__codigo__in=summary.affected_branches)
                .select_related("sucursal_relacionada")
                .order_by("id")
            )
            for project in projects:
                # Refresh stays scoped by affected branch/project, but not by a single month,
                # because payback/ROI/forecast/confidence depend on cumulative monthly history.
                self._refresh_service.refresh_project(project, until=refresh_until, user=user)
            summary.project_refresh_count = len(projects)
            summary.project_ids = [project.id for project in projects]

        log_event(
            user,
            "IMPORT",
            "reportes.GastoOperativoMensual",
            workbook.name,
            payload={
                "processed_rows": summary.processed_rows,
                "loaded_rows": summary.loaded_rows,
                "created": summary.created,
                "updated": summary.updated,
                "skipped_non_real": summary.skipped_non_real,
                "affected_branches": summary.affected_branches,
                "periods": summary.periods,
                "project_ids": summary.project_ids,
            },
        )
        return summary
