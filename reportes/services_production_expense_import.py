from __future__ import annotations

import unicodedata
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path

from django.db import transaction
from openpyxl import load_workbook

from reportes.models import CategoriaGasto, CentroCosto, GastoOperativoMensual


SPANISH_MONTHS = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}


def _normalize_key(value) -> str:
    raw = str(value or "").strip().upper()
    raw = unicodedata.normalize("NFKD", raw)
    raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
    return " ".join(raw.replace("/", " ").replace("-", " ").split())


def _to_decimal(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value).replace(",", "").strip())


PAYROLL_CATEGORY_MAP = {
    "SUELDO": "MANO_OBRA_PROD",
    "FESTIVOS": "MANO_OBRA_PROD",
    "VACACIONES": "MANO_OBRA_PROD",
    "PRIMA VACACIONES": "MANO_OBRA_PROD",
    "BONOS POR RESULTADOS": "MANO_OBRA_PROD",
    "BONOS PUNTUALIDAD": "MANO_OBRA_PROD",
    "BONOS ASISTENCIA": "MANO_OBRA_PROD",
    "IMSS": "MANO_OBRA_PROD",
    "INFONAVIT": "MANO_OBRA_PROD",
    "UTILIDADES": "MANO_OBRA_PROD",
    "AGUINALDO": "MANO_OBRA_PROD",
    "UNIFORMES": "INDIRECTO_PROD",
    "PLAYERA": "INDIRECTO_PROD",
    "POLO": "INDIRECTO_PROD",
    "MANDIL": "INDIRECTO_PROD",
    "GORRA": "INDIRECTO_PROD",
}

PRODUCTION_OVERHEAD_CATEGORY_MAP = {
    "AGUA POTABLE": "INDIRECTO_PROD",
    "ENERGIA ELECTRICA": "INDIRECTO_PROD",
    "GAS": "INDIRECTO_PROD",
    "ARTICULOS PARA ASEO Y LIMPIEZA": "INDIRECTO_PROD",
    "MATERIAL DE SEGUIRIDAD E HIGIENE": "INDIRECTO_PROD",
    "HERRAMIENTAS DE TRABAJO": "INDIRECTO_PROD",
    "FUMIGACION Y SANITIZACION": "INDIRECTO_PROD",
    "BATIDORA": "INDIRECTO_PROD",
    "MESA DE TRABAJO": "INDIRECTO_PROD",
    "MICROONDAS": "INDIRECTO_PROD",
    "MASCARILLA Y FILTROS": "INDIRECTO_PROD",
    "COMPRAS ESPECIALES": "INDIRECTO_PROD",
    "MANTENIMIENTO REMODELACION INSTALACIONES": "INDIRECTO_PROD",
    "MANTANIMIENTO EQUIPO MAQUINARIA": "INDIRECTO_PROD",
    "CAPACITACION": "INDIRECTO_PROD",
    "COEPRIS": "INDIRECTO_PROD",
    "DIVERSOS": "INDIRECTO_PROD",
    "AGUA PURIFICADA": "INDIRECTO_PROD",
}

PRODUCTION_PAYROLL_CATEGORY_MAP = {
    "SUELDO": "MANO_OBRA_PROD",
    "DIAS FESTIVOS": "MANO_OBRA_PROD",
    "VACACIONES": "MANO_OBRA_PROD",
    "PRIMA VACACIONAL": "MANO_OBRA_PROD",
    "BONO POR RESULTADOS": "MANO_OBRA_PROD",
    "BONO DE PUNTUALIDAD Y ASISTENCIA": "MANO_OBRA_PROD",
    "IMSS": "MANO_OBRA_PROD",
    "INFONAVIT": "MANO_OBRA_PROD",
    "AGUINALDO": "MANO_OBRA_PROD",
    "UTILIDADES": "MANO_OBRA_PROD",
    "PLAYERA": "MANO_OBRA_PROD",
    "GORRA": "MANO_OBRA_PROD",
    "MANDIL": "MANO_OBRA_PROD",
    "BLUSA": "MANO_OBRA_PROD",
}


@dataclass
class ProductionExpenseImportSummary:
    created: int = 0
    updated: int = 0
    deleted: int = 0
    periods: set[str] = field(default_factory=set)
    skipped_concepts: dict[str, set[str]] = field(default_factory=dict)
    flagged_outliers: list[dict[str, str]] = field(default_factory=list)

    def register_skip(self, source: str, concept: str):
        self.skipped_concepts.setdefault(source, set()).add(concept)

    def register_outlier(
        self,
        *,
        source: str,
        concept: str,
        period: date,
        budget: Decimal,
        actual: Decimal,
        ratio: Decimal,
    ) -> None:
        self.flagged_outliers.append(
            {
                "source": source,
                "concept": concept,
                "period": period.isoformat(),
                "budget": str(budget),
                "actual": str(actual),
                "ratio": str(ratio),
            }
        )


class ProductionExpenseImportService:
    PRODUCTION_FILENAME = "PRESUPUESTO PRODUCCIÓN 2026 AUTORIZADO.xlsx"
    PAYROLL_FILENAME = "PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx"
    PRODUCTION_SHEET = "PRESUPUESTO PRODUCCIÓN"
    PAYROLL_SHEET = "PRODUCCCION"
    OUTLIER_BUDGET_RATIO = Decimal("5")
    WATER_DECIMAL_SHIFT_RATIO = Decimal("5")
    WATER_DECIMAL_SHIFT_MIN_AMOUNT = Decimal("50000")

    def __init__(self):
        self.prod_center = CentroCosto.objects.get(codigo="PROD")
        self.categories = {
            categoria.codigo: categoria
            for categoria in CategoriaGasto.objects.filter(codigo__in={"MANO_OBRA_PROD", "INDIRECTO_PROD"})
        }

    def _upsert_expense(
        self,
        *,
        summary: ProductionExpenseImportSummary,
        seen_keys: dict[tuple[str, str], set[str]],
        source_file: str,
        source_sheet: str,
        source_row: int,
        period: date,
        concept: str,
        category_code: str,
        amount: Decimal,
    ) -> None:
        external_key = f"PROD|{source_file}|{source_sheet}|{source_row}|{period.isoformat()}|{category_code}"
        seen_keys.setdefault((source_file, source_sheet), set()).add(external_key)
        _, was_created = GastoOperativoMensual.objects.update_or_create(
            external_key=external_key,
            defaults={
                "periodo": period,
                "centro_costo": self.prod_center,
                "categoria_gasto": self.categories[category_code],
                "monto": amount,
                "tipo_dato": GastoOperativoMensual.TIPO_DATO_REAL,
                "fuente": GastoOperativoMensual.FUENTE_IMPORTADA,
                "es_estimado": False,
                "comentario": concept,
                "archivo_soporte": source_file,
            },
        )
        summary.created += int(was_created)
        summary.updated += int(not was_created)
        summary.periods.add(period.isoformat())

    def _cleanup_stale_rows(
        self,
        *,
        summary: ProductionExpenseImportSummary,
        source_file: str,
        source_sheet: str,
        seen_keys: set[str],
    ) -> None:
        stale_qs = GastoOperativoMensual.objects.filter(
            external_key__startswith=f"PROD|{source_file}|{source_sheet}|"
        )
        if seen_keys:
            stale_qs = stale_qs.exclude(external_key__in=seen_keys)
        deleted_count, _ = stale_qs.delete()
        summary.deleted += deleted_count

    def _parse_payroll_months(self, worksheet) -> list[tuple[int, date, int]]:
        result = []
        for col in range(1, worksheet.max_column + 1):
            header = _normalize_key(worksheet.cell(3, col).value)
            for month_name, month_number in SPANISH_MONTHS.items():
                if header == f"{month_name} REAL":
                    result.append((month_number, date(2026, month_number, 1), col - 1, col))
        return result

    def _parse_grouped_actual_months(self, worksheet) -> list[tuple[int, date, int, int]]:
        result = []
        current_month = None
        current_budget_col = None
        for col in range(1, worksheet.max_column + 1):
            top_label = _normalize_key(worksheet.cell(3, col).value)
            detail_label = _normalize_key(worksheet.cell(4, col).value)
            if top_label in SPANISH_MONTHS:
                current_month = top_label
                current_budget_col = None
            if current_month and detail_label == "PRESUPUESTADO":
                current_budget_col = col
            if current_month and detail_label == "REAL":
                month_number = SPANISH_MONTHS[current_month]
                result.append((month_number, date(2026, month_number, 1), current_budget_col or 0, col))
        return result

    def _register_outlier_if_needed(
        self,
        *,
        summary: ProductionExpenseImportSummary,
        source: str,
        concept: str,
        period: date,
        budget: Decimal,
        actual: Decimal,
    ) -> None:
        if budget <= 0 or actual <= 0:
            return
        ratio = actual / budget
        if ratio >= self.OUTLIER_BUDGET_RATIO:
            summary.register_outlier(
                source=source,
                concept=concept,
                period=period,
                budget=budget,
                actual=actual,
                ratio=ratio,
            )

    def _correct_amount_if_needed(
        self,
        *,
        summary: ProductionExpenseImportSummary,
        source: str,
        concept: str,
        period: date,
        budget: Decimal,
        amount: Decimal,
    ) -> Decimal:
        normalized_concept = _normalize_key(concept)
        if (
            normalized_concept == "AGUA POTABLE"
            and budget > 0
            and amount >= self.WATER_DECIMAL_SHIFT_MIN_AMOUNT
            and (amount / budget) >= self.WATER_DECIMAL_SHIFT_RATIO
        ):
            corrected = amount / Decimal("100")
            summary.register_outlier(
                source=source,
                concept=f"{concept} (corregido dos decimales)",
                period=period,
                budget=budget,
                actual=amount,
                ratio=amount / budget,
            )
            return corrected
        return amount

    def _import_payroll(
        self,
        workbook_path: Path,
        summary: ProductionExpenseImportSummary,
        seen_keys: dict[tuple[str, str], set[str]],
    ) -> None:
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        worksheet = workbook[self.PAYROLL_SHEET]
        actual_columns = self._parse_payroll_months(worksheet)
        for row_index in range(4, worksheet.max_row + 1):
            raw_concept = worksheet.cell(row_index, 1).value
            normalized_concept = _normalize_key(raw_concept)
            if not normalized_concept or normalized_concept == "TOTAL POR MES":
                continue
            category_code = PAYROLL_CATEGORY_MAP.get(normalized_concept)
            if not category_code:
                summary.register_skip(self.PAYROLL_SHEET, str(raw_concept or ""))
                continue
            for _, period, budget_col, actual_col in actual_columns:
                raw_amount = worksheet.cell(row_index, actual_col).value
                if raw_amount is None:
                    continue
                amount = _to_decimal(raw_amount)
                budget = _to_decimal(worksheet.cell(row_index, budget_col).value) if budget_col else Decimal("0")
                amount = self._correct_amount_if_needed(
                    summary=summary,
                    source=self.PAYROLL_SHEET,
                    concept=str(raw_concept or "").strip(),
                    period=period,
                    budget=budget,
                    amount=amount,
                )
                self._register_outlier_if_needed(
                    summary=summary,
                    source=self.PAYROLL_SHEET,
                    concept=str(raw_concept or "").strip(),
                    period=period,
                    budget=budget,
                    actual=amount,
                )
                self._upsert_expense(
                    summary=summary,
                    seen_keys=seen_keys,
                    source_file=workbook_path.name,
                    source_sheet=self.PAYROLL_SHEET,
                    source_row=row_index,
                    period=period,
                    concept=str(raw_concept or "").strip(),
                    category_code=category_code,
                    amount=amount,
                )

    def _import_production_overhead(
        self,
        workbook_path: Path,
        summary: ProductionExpenseImportSummary,
        seen_keys: dict[tuple[str, str], set[str]],
        *,
        include_payroll_from_production: bool = False,
        through_month: int | None = None,
    ) -> None:
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        worksheet = workbook[self.PRODUCTION_SHEET]
        actual_columns = self._parse_grouped_actual_months(worksheet)
        for row_index in range(5, worksheet.max_row + 1):
            raw_concept = worksheet.cell(row_index, 2).value
            normalized_concept = _normalize_key(raw_concept)
            if not normalized_concept:
                continue
            category_code = PRODUCTION_OVERHEAD_CATEGORY_MAP.get(normalized_concept)
            if category_code is None and include_payroll_from_production:
                category_code = PRODUCTION_PAYROLL_CATEGORY_MAP.get(normalized_concept)
            if not category_code:
                continue
            for _, period, budget_col, actual_col in actual_columns:
                if through_month is not None and period.month > through_month:
                    continue
                raw_amount = worksheet.cell(row_index, actual_col).value
                if raw_amount is None:
                    continue
                amount = _to_decimal(raw_amount)
                budget = _to_decimal(worksheet.cell(row_index, budget_col).value) if budget_col else Decimal("0")
                amount = self._correct_amount_if_needed(
                    summary=summary,
                    source=self.PRODUCTION_SHEET,
                    concept=str(raw_concept or "").strip(),
                    period=period,
                    budget=budget,
                    amount=amount,
                )
                self._register_outlier_if_needed(
                    summary=summary,
                    source=self.PRODUCTION_SHEET,
                    concept=str(raw_concept or "").strip(),
                    period=period,
                    budget=budget,
                    actual=amount,
                )
                self._upsert_expense(
                    summary=summary,
                    seen_keys=seen_keys,
                    source_file=workbook_path.name,
                    source_sheet=self.PRODUCTION_SHEET,
                    source_row=row_index,
                    period=period,
                    concept=str(raw_concept or "").strip(),
                    category_code=category_code,
                    amount=amount,
                )

    @transaction.atomic
    def import_production_workbook(
        self,
        workbook_path: str | Path,
        *,
        through_month: int | None = None,
    ) -> ProductionExpenseImportSummary:
        path = Path(workbook_path).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(path)
        summary = ProductionExpenseImportSummary()
        seen_keys: dict[tuple[str, str], set[str]] = {}
        self._import_production_overhead(
            path,
            summary,
            seen_keys,
            include_payroll_from_production=True,
            through_month=through_month,
        )
        for (source_file, source_sheet), keys in seen_keys.items():
            self._cleanup_stale_rows(summary=summary, source_file=source_file, source_sheet=source_sheet, seen_keys=keys)
        return summary

    @transaction.atomic
    def import_folder(self, folder_path: str | Path) -> ProductionExpenseImportSummary:
        folder = Path(folder_path).expanduser().resolve()
        if not folder.exists():
            raise FileNotFoundError(folder)
        summary = ProductionExpenseImportSummary()
        production_path = folder / self.PRODUCTION_FILENAME
        payroll_path = folder / self.PAYROLL_FILENAME
        if not production_path.exists():
            raise FileNotFoundError(production_path)
        if not payroll_path.exists():
            raise FileNotFoundError(payroll_path)
        seen_keys: dict[tuple[str, str], set[str]] = {}

        self._import_production_overhead(production_path, summary, seen_keys)
        self._import_payroll(payroll_path, summary, seen_keys)
        for (source_file, source_sheet), keys in seen_keys.items():
            self._cleanup_stale_rows(summary=summary, source_file=source_file, source_sheet=source_sheet, seen_keys=keys)
        return summary
