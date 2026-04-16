from __future__ import annotations

import unicodedata
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path

from django.db import transaction
from openpyxl import load_workbook

from core.models import Sucursal
from reportes.models import CategoriaGasto, CentroCosto, GastoOperativoMensual


SPANISH_MONTHS = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}


def _normalize_key(value) -> str:
    raw = str(value or "").strip().upper()
    raw = unicodedata.normalize("NFKD", raw)
    raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
    return " ".join(raw.replace("/", " ").replace("-", " ").replace(".", " ").replace(",", " ").split())


def _to_decimal(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value).replace(",", "").strip())


BRANCH_PAYROLL_CONCEPTS = {
    "SUELDO",
    "DIAS FESTIVOS",
    "VACACIONES",
    "PRIMA VACACIONAL",
    "AGUINALDO",
    "IMSS",
    "INFONAVIT",
    "BONO POR RESULTADOS",
}
BRANCH_RENT_CONCEPTS = {"ARRENDAMIENTO LOCAL"}
BRANCH_PLATFORM_CONCEPTS = {
    "CAJA REGISTRADORA",
    "CAMARAS",
    "COMPUTADORA",
    "IMPRESORA",
    "TELEFONO",
    "TELEFONO E INTERNET",
    "TELEVISION",
}
BRANCH_INDIRECTO_CONCEPTS = {
    "AGUA POTABLE",
    "AGUA PURIFICADA",
    "ARTICULOS PARA ASEO Y LIMPIEZA",
    "DECORACION",
    "ENERGIA ELECTRICA",
    "ETIQUETAS BOLSAS CAJAS Y EMPAQUES",
    "FUMIGACION Y SANITIZACION",
    "GASTOS DIVERSOS VENTAS",
    "HERRAMIENTAS DE TRABAJO",
    "MANDIL",
    "MATERIAL DE SEGUIRIDAD E HIGIENE",
    "MATERIAL IMPRESO",
    "MENUS",
    "MESA FRIA",
    "MINISPLIT",
    "MANTENIMIENTO EQUIPO MAQUINARIA",
    "MANTENIMIENTO REMODELACION INSTALACIONES",
    "PAPELERIA",
    "PATROCINIOS",
    "PLAYERA",
    "PUBLICIDAD",
    "REFIGERADOR 1",
    "REFIGERADOR 2",
    "REFRIGERADOR",
    "SERVICIOS PUBLICOS",
    "UNIFORMES",
    "VITRINA",
    "VITRINAS",
    "APERTURA SUCURSAL",
    "ADQUISICION DE EQUIPO MAQUINARIA",
    "GORRA",
}

ADMIN_SYSTEMS_CONCEPTS = {
    "CONTPAQ",
    "POINT",
    "LICENCIAS Y SERVICIOS DE SISTEMAS",
    "LINEAS SRA POLYANA TELCEL",
    "SERVICIOS DE CELULAR",
    "TELEFONO E INTERNET",
}
ADMIN_SKIP_CONCEPTS = {
    "INGRESOS",
    "EGRESOS",
    "COSTOS",
    "UTILIDAD BRUTA",
    "UTILIDAD O PERDIDA",
    "VENTA COMPLEMENTOS",
    "VENTA POSTRES",
}
BIMONTHLY_TOLERANT_CONCEPTS = {
    "AGUA POTABLE",
    "ENERGIA ELECTRICA",
    "SERVICIOS PUBLICOS",
    "TELEFONO",
    "TELEFONO E INTERNET",
}


@dataclass
class BranchAdminExpenseImportSummary:
    processed_rows: int = 0
    created: int = 0
    updated: int = 0
    deleted: int = 0
    periods: set[str] = field(default_factory=set)
    affected_branches: set[str] = field(default_factory=set)
    skipped_concepts: dict[str, set[str]] = field(default_factory=dict)
    flagged_outliers: list[dict[str, str]] = field(default_factory=list)

    def register_skip(self, source: str, concept: str):
        self.skipped_concepts.setdefault(source, set()).add(concept)

    def register_outlier(
        self,
        *,
        source: str,
        concept: str,
        period: date,
        budget: Decimal,
        actual: Decimal,
        ratio: Decimal,
    ):
        self.flagged_outliers.append(
            {
                "source": source,
                "concept": concept,
                "period": period.isoformat(),
                "budget": str(budget),
                "actual": str(actual),
                "ratio": str(ratio),
            }
        )


class BranchAdminExpenseImportService:
    SALES_FILENAME = "PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx"
    ADMIN_FILENAME = "PRESUPUESTO ADMINISTRACIÓN 2026.xlsx"
    ADMIN_SHEET = "ADMON"
    OUTLIER_BUDGET_RATIO = Decimal("5")
    MAX_SCAN_COLUMNS = 60

    def __init__(
        self,
        *,
        target_year: int = 2026,
        branch_tipo_dato: str = GastoOperativoMensual.TIPO_DATO_REAL,
        admin_tipo_dato: str = GastoOperativoMensual.TIPO_DATO_REAL,
        external_prefix: str = "OPEX",
    ):
        self.target_year = int(target_year)
        self.branch_tipo_dato = branch_tipo_dato
        self.admin_tipo_dato = admin_tipo_dato
        self.external_prefix = external_prefix
        self.categories = {
            categoria.codigo: categoria
            for categoria in CategoriaGasto.objects.filter(
                codigo__in={"NOMINA_SUC", "RENTA_SUC", "INDIRECTO_SUC", "PLATAFORMAS", "ADMIN_CORP", "SISTEMAS_CORP"}
            )
        }
        self.branch_centers = self._build_branch_centers()
        self.corp_center = CentroCosto.objects.get(codigo="CORP")

    def _build_branch_centers(self) -> dict[str, CentroCosto]:
        centers = {}
        for centro in CentroCosto.objects.filter(codigo__startswith="SUC_").select_related("sucursal"):
            keys = {_normalize_key(centro.codigo.replace("SUC_", ""))}
            if centro.sucursal_id:
                keys.add(_normalize_key(centro.sucursal.codigo))
                keys.add(_normalize_key(centro.sucursal.nombre))
            for key in keys:
                centers[key] = centro
        aliases = {
            "EL TUNEL": "SUC_EL_TUNEL",
            "PLAZA NIO": "SUC_PLAZA_NIO",
            "GLORIAS": "SUC_LAS_GLORIAS",
        }
        for alias, center_code in aliases.items():
            centro = CentroCosto.objects.filter(codigo=center_code).first()
            if centro is not None:
                centers[_normalize_key(alias)] = centro
        return centers

    def _parse_actual_columns(self, worksheet) -> list[tuple[date, int, int]]:
        result = []
        current_month = None
        current_budget_col = None
        scan_limit = min(worksheet.max_column, self.MAX_SCAN_COLUMNS)
        for col in range(1, scan_limit + 1):
            top_label = _normalize_key(worksheet.cell(3, col).value)
            detail_label = _normalize_key(worksheet.cell(4, col).value)
            if top_label in SPANISH_MONTHS:
                current_month = top_label
                current_budget_col = None
            if current_month and detail_label == "PRESUPUESTADO":
                current_budget_col = col
            if current_month and detail_label == "REAL":
                month_number = SPANISH_MONTHS[current_month]
                result.append((date(self.target_year, month_number, 1), current_budget_col or 0, col))
        return result

    def _register_outlier_if_needed(
        self,
        *,
        summary: BranchAdminExpenseImportSummary,
        source: str,
        concept: str,
        period: date,
        budget: Decimal,
        actual: Decimal,
    ) -> None:
        normalized = _normalize_key(concept)
        if normalized in BIMONTHLY_TOLERANT_CONCEPTS:
            return
        if budget <= 0 or actual <= 0:
            return
        ratio = actual / budget
        if ratio >= self.OUTLIER_BUDGET_RATIO:
            summary.register_outlier(
                source=source,
                concept=concept,
                period=period,
                budget=budget,
                actual=actual,
                ratio=ratio,
            )

    def _cleanup_stale_rows(
        self,
        *,
        summary: BranchAdminExpenseImportSummary,
        source_file: str,
        source_sheet: str,
        seen_keys: set[str],
    ) -> None:
        stale_qs = GastoOperativoMensual.objects.filter(
            external_key__startswith=f"OPEX|{source_file}|{source_sheet}|"
        )
        if seen_keys:
            stale_qs = stale_qs.exclude(external_key__in=seen_keys)
        deleted_count, _ = stale_qs.delete()
        summary.deleted += deleted_count

    def _upsert_expense(
        self,
        *,
        summary: BranchAdminExpenseImportSummary,
        seen_keys: dict[tuple[str, str], set[str]],
        prefix: str,
        source_file: str,
        source_sheet: str,
        source_row: int,
        period: date,
        center: CentroCosto,
        category_code: str,
        concept: str,
        amount: Decimal,
        tipo_dato: str,
    ) -> None:
        external_key = f"{prefix}|{source_file}|{source_sheet}|{source_row}|{period.isoformat()}|{category_code}"
        seen_keys.setdefault((source_file, source_sheet), set()).add(external_key)
        _, was_created = GastoOperativoMensual.objects.update_or_create(
            external_key=external_key,
            defaults={
                "periodo": period,
                "centro_costo": center,
                "categoria_gasto": self.categories[category_code],
                "monto": amount,
                "tipo_dato": tipo_dato,
                "fuente": GastoOperativoMensual.FUENTE_IMPORTADA,
                "es_estimado": False,
                "comentario": concept,
                "archivo_soporte": source_file,
            },
        )
        summary.processed_rows += 1
        summary.created += int(was_created)
        summary.updated += int(not was_created)
        summary.periods.add(period.isoformat())

    def _resolve_branch_category(self, concept: str) -> str | None:
        normalized = _normalize_key(concept)
        if normalized in BRANCH_PAYROLL_CONCEPTS:
            return "NOMINA_SUC"
        if normalized in BRANCH_RENT_CONCEPTS:
            return "RENTA_SUC"
        if normalized in BRANCH_PLATFORM_CONCEPTS:
            return "PLATAFORMAS"
        if normalized in BRANCH_INDIRECTO_CONCEPTS:
            return "INDIRECTO_SUC"
        return None

    def _resolve_admin_category(self, concept: str) -> str | None:
        normalized = _normalize_key(concept)
        if normalized in ADMIN_SKIP_CONCEPTS:
            return None
        if normalized in ADMIN_SYSTEMS_CONCEPTS:
            return "SISTEMAS_CORP"
        return "ADMIN_CORP"

    def _import_sales_workbook(self, workbook_path: Path, summary: BranchAdminExpenseImportSummary, seen_keys: dict[tuple[str, str], set[str]]) -> None:
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        ignored_sheets = {"GENERAL", "LISTA GASTOS"}
        for sheet_name in workbook.sheetnames:
            if sheet_name in ignored_sheets:
                continue
            center = self.branch_centers.get(_normalize_key(sheet_name))
            if center is None:
                summary.register_skip("SHEETS", sheet_name)
                continue
            if center.sucursal_id:
                summary.affected_branches.add(center.sucursal.codigo)
            worksheet = workbook[sheet_name]
            actual_columns = self._parse_actual_columns(worksheet)
            for row_index in range(5, worksheet.max_row + 1):
                raw_concept = worksheet.cell(row_index, 2).value
                concept = str(raw_concept or "").strip()
                if not concept:
                    continue
                category_code = self._resolve_branch_category(concept)
                if category_code is None:
                    summary.register_skip(sheet_name, concept)
                    continue
                for period, budget_col, actual_col in actual_columns:
                    raw_amount = worksheet.cell(row_index, actual_col).value
                    if raw_amount is None:
                        continue
                    amount = _to_decimal(raw_amount)
                    budget = _to_decimal(worksheet.cell(row_index, budget_col).value) if budget_col else Decimal("0")
                    self._register_outlier_if_needed(
                        summary=summary,
                        source=sheet_name,
                        concept=concept,
                        period=period,
                        budget=budget,
                        actual=amount,
                    )
                    self._upsert_expense(
                        summary=summary,
                        seen_keys=seen_keys,
                        prefix=self.external_prefix,
                        source_file=workbook_path.name,
                        source_sheet=sheet_name,
                        source_row=row_index,
                        period=period,
                        center=center,
                        category_code=category_code,
                        concept=concept,
                        amount=amount,
                        tipo_dato=self.branch_tipo_dato,
                    )

    def _import_admin_workbook(self, workbook_path: Path, summary: BranchAdminExpenseImportSummary, seen_keys: dict[tuple[str, str], set[str]]) -> None:
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        worksheet = workbook[self.ADMIN_SHEET]
        actual_columns = self._parse_actual_columns(worksheet)
        for row_index in range(5, worksheet.max_row + 1):
            raw_concept = worksheet.cell(row_index, 2).value
            concept = str(raw_concept or "").strip()
            if not concept:
                continue
            category_code = self._resolve_admin_category(concept)
            if category_code is None:
                continue
            for period, budget_col, actual_col in actual_columns:
                raw_amount = worksheet.cell(row_index, actual_col).value
                if raw_amount is None:
                    continue
                amount = _to_decimal(raw_amount)
                budget = _to_decimal(worksheet.cell(row_index, budget_col).value) if budget_col else Decimal("0")
                self._register_outlier_if_needed(
                    summary=summary,
                    source=self.ADMIN_SHEET,
                    concept=concept,
                    period=period,
                    budget=budget,
                    actual=amount,
                )
                self._upsert_expense(
                    summary=summary,
                    seen_keys=seen_keys,
                    prefix=self.external_prefix,
                    source_file=workbook_path.name,
                    source_sheet=self.ADMIN_SHEET,
                    source_row=row_index,
                    period=period,
                    center=self.corp_center,
                    category_code=category_code,
                    concept=concept,
                    amount=amount,
                    tipo_dato=self.admin_tipo_dato,
                )

    @transaction.atomic
    def import_sales_workbook(self, workbook_path: str | Path) -> BranchAdminExpenseImportSummary:
        path = Path(workbook_path).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(path)
        summary = BranchAdminExpenseImportSummary()
        seen_keys: dict[tuple[str, str], set[str]] = {}
        self._import_sales_workbook(path, summary, seen_keys)
        for (source_file, source_sheet), keys in seen_keys.items():
            self._cleanup_stale_rows(summary=summary, source_file=source_file, source_sheet=source_sheet, seen_keys=keys)
        return summary

    @transaction.atomic
    def import_folder(self, folder_path: str | Path) -> BranchAdminExpenseImportSummary:
        folder = Path(folder_path).expanduser().resolve()
        if not folder.exists():
            raise FileNotFoundError(folder)
        sales_path = folder / self.SALES_FILENAME
        admin_path = folder / self.ADMIN_FILENAME
        if not sales_path.exists():
            raise FileNotFoundError(sales_path)
        if not admin_path.exists():
            raise FileNotFoundError(admin_path)

        summary = BranchAdminExpenseImportSummary()
        seen_keys: dict[tuple[str, str], set[str]] = {}
        self._import_sales_workbook(sales_path, summary, seen_keys)
        self._import_admin_workbook(admin_path, summary, seen_keys)
        for (source_file, source_sheet), keys in seen_keys.items():
            self._cleanup_stale_rows(summary=summary, source_file=source_file, source_sheet=source_sheet, seen_keys=keys)
        return summary
