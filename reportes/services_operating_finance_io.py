from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path

from django.db import transaction
from openpyxl import Workbook, load_workbook

from reportes.models import CategoriaGasto, CentroCosto, GastoOperativoMensual


def _normalize_text(value) -> str:
    return str(value or "").strip()


def _parse_period(raw_value) -> date:
    value = _normalize_text(raw_value)
    if not value:
        raise ValueError("periodo vacío")
    try:
        return date.fromisoformat(f"{value}-01")
    except ValueError as exc:
        raise ValueError(f"periodo inválido: {value}") from exc


def _parse_decimal(raw_value) -> Decimal:
    value = _normalize_text(raw_value)
    if not value:
        return Decimal("0")
    normalized = value.replace(",", "")
    return Decimal(normalized)


def _parse_bool(raw_value) -> bool:
    value = _normalize_text(raw_value).lower()
    return value in {"1", "true", "si", "sí", "yes", "x"}


@dataclass
class OperatingExpenseImportSummary:
    created: int
    updated: int
    periods: list[str]


class OperatingFinanceTemplateService:
    HEADER = [
        "external_key",
        "periodo",
        "centro_costo",
        "categoria_gasto",
        "monto",
        "tipo_dato",
        "fuente",
        "es_estimado",
        "comentario",
        "archivo_soporte",
    ]

    def export_template(self, output_path: str | Path) -> Path:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Gastos"
        ws.append(self.HEADER)

        catalog_sheet = wb.create_sheet("Catalogos")
        catalog_sheet.append(["tipo", "codigo", "nombre", "detalle"])
        for centro in CentroCosto.objects.order_by("codigo"):
            catalog_sheet.append(
                [
                    "CENTRO_COSTO",
                    centro.codigo,
                    centro.nombre,
                    centro.sucursal.codigo if centro.sucursal_id else centro.tipo,
                ]
            )
        for categoria in CategoriaGasto.objects.order_by("codigo"):
            catalog_sheet.append(
                [
                    "CATEGORIA_GASTO",
                    categoria.codigo,
                    categoria.nombre,
                    categoria.capa_objetivo,
                ]
            )

        rules_sheet = wb.create_sheet("Reglas")
        rules_sheet.append(["campo", "descripcion"])
        rules_sheet.append(["external_key", "Clave única por fila para reimportar sin duplicar."])
        rules_sheet.append(["periodo", "Formato YYYY-MM."])
        rules_sheet.append(["centro_costo", "Usar código de la hoja Catalogos."])
        rules_sheet.append(["categoria_gasto", "Usar código de la hoja Catalogos."])
        rules_sheet.append(["monto", "Monto mensual en moneda local."])
        rules_sheet.append(["tipo_dato", "REAL o PRESUPUESTO."])
        rules_sheet.append(["fuente", "MANUAL o IMPORTADA."])
        rules_sheet.append(["es_estimado", "1/0, Sí/No, True/False."])
        rules_sheet.append(["comentario", "Referencia corta del gasto."])
        rules_sheet.append(["archivo_soporte", "Nombre o ruta del soporte."])
        rules_sheet.append(
            [
                "ejemplo",
                "CEDIS_MO_2026_03 | 2026-03 | PROD | MANO_OBRA_PROD | 15000.00 | REAL | IMPORTADA | 0 | Nómina producción marzo | nomina_marzo.xlsx",
            ]
        )

        wb.save(output)
        return output


class OperatingFinanceExpenseImportService:
    REQUIRED_HEADERS = {
        "external_key",
        "periodo",
        "centro_costo",
        "categoria_gasto",
        "monto",
    }

    @transaction.atomic
    def import_workbook(self, workbook_path: str | Path) -> OperatingExpenseImportSummary:
        workbook = load_workbook(filename=workbook_path, data_only=True)
        if "Gastos" not in workbook.sheetnames:
            raise ValueError("El archivo debe contener una hoja llamada 'Gastos'.")

        ws = workbook["Gastos"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("La hoja 'Gastos' está vacía.")
        headers = [_normalize_text(value) for value in rows[0]]
        missing = self.REQUIRED_HEADERS - set(headers)
        if missing:
            raise ValueError(f"Faltan columnas requeridas: {', '.join(sorted(missing))}.")

        header_map = {header: index for index, header in enumerate(headers)}
        created = 0
        updated = 0
        periods: set[str] = set()

        for line_number, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue
            external_key = _normalize_text(row[header_map["external_key"]])
            if not external_key:
                raise ValueError(f"Fila {line_number}: external_key es obligatorio.")
            period = _parse_period(row[header_map["periodo"]])
            center_code = _normalize_text(row[header_map["centro_costo"]])
            category_code = _normalize_text(row[header_map["categoria_gasto"]])
            if not center_code or not category_code:
                raise ValueError(f"Fila {line_number}: centro_costo y categoria_gasto son obligatorios.")

            centro = CentroCosto.objects.filter(codigo=center_code).first()
            if centro is None:
                raise ValueError(f"Fila {line_number}: centro_costo desconocido '{center_code}'.")
            categoria = CategoriaGasto.objects.filter(codigo=category_code).first()
            if categoria is None:
                raise ValueError(f"Fila {line_number}: categoria_gasto desconocida '{category_code}'.")

            monto = _parse_decimal(row[header_map["monto"]])
            tipo_dato = _normalize_text(row[header_map.get("tipo_dato", -1)] if "tipo_dato" in header_map else "")
            if tipo_dato not in {
                GastoOperativoMensual.TIPO_DATO_REAL,
                GastoOperativoMensual.TIPO_DATO_PRESUPUESTO,
            }:
                tipo_dato = GastoOperativoMensual.TIPO_DATO_REAL
            fuente = _normalize_text(row[header_map.get("fuente", -1)] if "fuente" in header_map else "")
            if fuente not in {GastoOperativoMensual.FUENTE_MANUAL, GastoOperativoMensual.FUENTE_IMPORTADA}:
                fuente = GastoOperativoMensual.FUENTE_IMPORTADA
            es_estimado = _parse_bool(row[header_map.get("es_estimado", -1)] if "es_estimado" in header_map else "")
            comentario = _normalize_text(row[header_map.get("comentario", -1)] if "comentario" in header_map else "")
            archivo_soporte = _normalize_text(
                row[header_map.get("archivo_soporte", -1)] if "archivo_soporte" in header_map else ""
            )

            _, was_created = GastoOperativoMensual.objects.update_or_create(
                external_key=external_key,
                defaults={
                    "periodo": period,
                    "centro_costo": centro,
                    "categoria_gasto": categoria,
                    "monto": monto,
                    "tipo_dato": tipo_dato,
                    "fuente": fuente,
                    "es_estimado": es_estimado,
                    "comentario": comentario,
                    "archivo_soporte": archivo_soporte,
                },
            )
            created += int(was_created)
            updated += int(not was_created)
            periods.add(period.isoformat())

        return OperatingExpenseImportSummary(
            created=created,
            updated=updated,
            periods=sorted(periods),
        )
