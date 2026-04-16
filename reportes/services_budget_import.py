from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path

from django.db import transaction
from openpyxl import load_workbook

from reportes.models import PresupuestoImport, PresupuestoLineaMensual


SPANISH_MONTHS = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}


def _clean(value) -> str:
    return str(value or "").strip()


def _to_decimal(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    raw = str(value).strip()
    if not raw:
        return Decimal("0")
    if raw.endswith("%"):
        try:
            return Decimal(raw[:-1]) / Decimal("100")
        except Exception:
            return Decimal("0")
    try:
        return Decimal(raw)
    except Exception:
        return Decimal("0")


def _safe_ratio(numerator: Decimal, denominator: Decimal) -> Decimal:
    if denominator == 0:
        return Decimal("0")
    return numerator / denominator


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


@dataclass
class GeneralBudgetImportSummary:
    imports_created: int
    imports_updated: int
    lines_created: int
    lines_updated: int
    periods: list[str]


class GeneralBudgetImportService:
    SHEET_NAME = "GENERAL"

    def _effective_max_col(self, worksheet) -> int:
        max_seen = 0
        empty_streak = 0
        limit = worksheet.max_column
        for col in range(1, limit + 1):
            row3 = _clean(worksheet.cell(3, col).value)
            row4 = _clean(worksheet.cell(4, col).value)
            if row3 or row4:
                max_seen = col
                empty_streak = 0
            else:
                empty_streak += 1
                if max_seen and empty_streak >= 24:
                    break
        return max_seen or min(limit, 64)

    def _find_title(self, worksheet) -> str:
        for row in range(1, 5):
            for col in range(1, 6):
                value = _clean(worksheet.cell(row, col).value)
                if value and "PRESUPUESTO" in value.upper():
                    return value
        return worksheet.title

    def _resolve_year(self, path: Path, title: str) -> int:
        text = f"{path.name} {title}".upper()
        match = re.search(r"\b(20\d{2})\b", text)
        if match:
            return int(match.group(1))
        raise ValueError(f"No pude resolver el año del presupuesto en {path.name}.")

    def _parse_layout(self, worksheet) -> tuple[dict[str, int], list[tuple[int, dict[str, int]]], int]:
        annual_row = 3
        detail_row = 4
        max_col = self._effective_max_col(worksheet)

        annual_cols: dict[str, int] = {}
        monthly_groups: list[tuple[int, date, dict[str, int]]] = []
        current_month = None
        for col in range(1, max_col + 1):
            top_label = _clean(worksheet.cell(annual_row, col).value).upper()
            detail_label = _clean(worksheet.cell(detail_row, col).value).upper()
            if top_label == "TOTAL ANUAL":
                if detail_label == "PRESUPUESTO":
                    annual_cols["budget"] = col
                elif detail_label == "RESULTADO":
                    annual_cols["actual"] = col
                elif "VARI" in detail_label:
                    annual_cols["variance"] = col
                continue
            if top_label in SPANISH_MONTHS:
                current_month = top_label
            if current_month and detail_label:
                monthly_groups.append((col, current_month, detail_label))

        grouped: dict[str, dict[str, int]] = {}
        for col, month_name, detail_label in monthly_groups:
            group = grouped.setdefault(month_name, {})
            if detail_label == "PRESUPUESTADO":
                group["budget"] = col
            elif detail_label == "REAL":
                group["actual"] = col
            elif "VARI" in detail_label:
                group["variance"] = col

        result = []
        for month_name, columns in grouped.items():
            if "budget" not in columns and "actual" not in columns:
                continue
            result.append((SPANISH_MONTHS[month_name], columns))
        if result:
            return annual_cols, result, 5

        flat_headers = {}
        for col in range(1, max_col + 1):
            header = _clean(worksheet.cell(3, col).value).upper()
            for month_name, month_number in SPANISH_MONTHS.items():
                if header.startswith(month_name):
                    if "PRESUPUEST" in header:
                        flat_headers.setdefault(month_number, {})["budget"] = col
                    elif "REAL" in header or "RESULT" in header:
                        flat_headers.setdefault(month_number, {})["actual"] = col
                    elif "VARI" in header:
                        flat_headers.setdefault(month_number, {})["variance"] = col
        flat_result = [(month_number, columns) for month_number, columns in flat_headers.items() if columns]
        return annual_cols, flat_result, 4

    @transaction.atomic
    def import_workbook(self, workbook_path: str | Path) -> GeneralBudgetImportSummary:
        path = Path(workbook_path).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(path)

        wb_formula = load_workbook(path, read_only=True, data_only=False)
        if self.SHEET_NAME not in wb_formula.sheetnames:
            raise ValueError(f"{path.name} no contiene hoja '{self.SHEET_NAME}'.")
        wb_values = load_workbook(path, read_only=True, data_only=True)
        ws_formula = wb_formula[self.SHEET_NAME]
        ws_values = wb_values[self.SHEET_NAME]

        title = self._find_title(ws_formula)
        year = self._resolve_year(path, title)
        annual_cols, monthly_templates, data_start_row = self._parse_layout(ws_formula)
        monthly_groups = [(month_number, date(year, month_number, 1), cols) for month_number, cols in monthly_templates]
        if not monthly_groups:
            raise ValueError(f"No pude encontrar bloques mensuales válidos en {path.name}.")

        file_hash = _sha256(path)
        import_obj, created = PresupuestoImport.objects.update_or_create(
            tipo=PresupuestoImport.TIPO_GENERAL,
            fuente_nombre=path.name,
            sheet_name=self.SHEET_NAME,
            defaults={
                "archivo_ruta": str(path),
                "archivo_hash": file_hash,
                "titulo": title,
                "metadata": {"year": year},
            },
        )

        periods: set[str] = set()
        lines_created = 0
        lines_updated = 0
        started = False
        for row_index in range(data_start_row, ws_values.max_row + 1):
            raw_account = ws_values.cell(row_index, 1).value
            raw_concept = ws_values.cell(row_index, 2).value
            account_code = _clean(raw_account)
            concept = _clean(raw_concept)
            if not isinstance(raw_concept, str) and account_code:
                concept = account_code
                account_code = ""
            elif not concept and account_code:
                concept = account_code
                account_code = ""
            if not concept:
                if started:
                    continue
                continue
            started = True

            annual_budget = _to_decimal(ws_values.cell(row_index, annual_cols.get("budget", 0)).value if annual_cols.get("budget") else 0)
            annual_actual = _to_decimal(ws_values.cell(row_index, annual_cols.get("actual", 0)).value if annual_cols.get("actual") else 0)
            annual_variance = _to_decimal(
                ws_values.cell(row_index, annual_cols.get("variance", 0)).value if annual_cols.get("variance") else 0
            )
            if annual_variance == 0 and annual_budget > 0 and annual_actual > 0:
                annual_variance = _safe_ratio(annual_actual, annual_budget)

            for month_number, period_start, columns in monthly_groups:
                monthly_budget = _to_decimal(ws_values.cell(row_index, columns.get("budget", 0)).value if columns.get("budget") else 0)
                monthly_actual = _to_decimal(ws_values.cell(row_index, columns.get("actual", 0)).value if columns.get("actual") else 0)
                monthly_variance = _to_decimal(
                    ws_values.cell(row_index, columns.get("variance", 0)).value if columns.get("variance") else 0
                )
                if monthly_variance == 0 and monthly_budget > 0 and monthly_actual > 0:
                    monthly_variance = _safe_ratio(monthly_actual, monthly_budget)
                if (
                    annual_budget == 0
                    and annual_actual == 0
                    and monthly_budget == 0
                    and monthly_actual == 0
                    and monthly_variance == 0
                    and annual_variance == 0
                ):
                    continue
                external_key = f"{path.name}|{self.SHEET_NAME}|{row_index}|{period_start.isoformat()}"
                _, was_created = PresupuestoLineaMensual.objects.update_or_create(
                    external_key=external_key,
                    defaults={
                        "importacion": import_obj,
                        "period": period_start,
                        "account_code": account_code,
                        "concept": concept,
                        "annual_budget": annual_budget,
                        "annual_actual": annual_actual,
                        "annual_variance": annual_variance,
                        "monthly_budget": monthly_budget,
                        "monthly_actual": monthly_actual,
                        "monthly_variance": monthly_variance,
                        "row_index": row_index,
                        "metadata": {
                            "month_number": month_number,
                            "source_title": title,
                        },
                    },
                )
                lines_created += int(was_created)
                lines_updated += int(not was_created)
                periods.add(period_start.isoformat())

        return GeneralBudgetImportSummary(
            imports_created=int(created),
            imports_updated=int(not created),
            lines_created=lines_created,
            lines_updated=lines_updated,
            periods=sorted(periods),
        )

    def import_folder(self, folder_path: str | Path) -> GeneralBudgetImportSummary:
        folder = Path(folder_path).expanduser().resolve()
        if not folder.exists():
            raise FileNotFoundError(folder)

        totals = GeneralBudgetImportSummary(0, 0, 0, 0, [])
        periods: set[str] = set()
        for path in sorted(folder.glob("*.xlsx")):
            try:
                summary = self.import_workbook(path)
            except ValueError:
                continue
            totals.imports_created += summary.imports_created
            totals.imports_updated += summary.imports_updated
            totals.lines_created += summary.lines_created
            totals.lines_updated += summary.lines_updated
            periods.update(summary.periods)
        totals.periods = sorted(periods)
        return totals
