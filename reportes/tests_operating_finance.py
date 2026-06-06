from __future__ import annotations

import json
from io import StringIO
from datetime import date
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from django.contrib.auth.models import Group, User
from django.core.exceptions import ImproperlyConfigured
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.db import IntegrityError
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook
from openpyxl import Workbook

from core.models import AuditLog, Sucursal
from maestros.models import CostoInsumo, Insumo, InsumoAlias, UnidadMedida
from pos_bridge.models import PointBranch, PointDailySale, PointMonthlySalesOfficial, PointProduct
from recetas.models import LineaReceta, Receta, RecetaCostoSemanal, RecetaPresentacionDerivada
from reportes.models import (
    AreaPresupuesto,
    CargaGastoOperativoArchivo,
    CategoriaGasto,
    CentroCosto,
    EmpresaResultadoMensual,
    FactVentaDiaria,
    GastoOperativoMensual,
    InsumoCostoHistoricoMensual,
    LineaPresupuestoMensual,
    ProductBusinessRule,
    PresupuestoImport,
    PresupuestoLineaMensual,
    PresupuestoResumenMensual,
    ProyectoInversion,
    ProyectoInversionSnapshotMensual,
    ProductoCostoOperativoMensual,
    ProductoPricingDecisionMensual,
    ProductoReventaCosto,
    ProductoReventaCostoHistoricoMensual,
    ProductoSucursalContribucionMensual,
    RecetaCostoHistoricoMensual,
    ReglaCostoHistoricoInsumo,
    ReglaAsignacionGasto,
    RubroPresupuesto,
)
from reportes.checks import (
    assert_critical_product_business_rules_present,
    collect_critical_product_business_rule_issues,
)
from reportes.product_business_rules import CRITICAL_FIXED_REVENTA_PRODUCT_NAMES
from reportes.services_historical_costing import MonthlyHistoricalCostingService
from reportes.services_operating_finance import (
    OperatingFinanceBootstrapService,
    OperatingFinanceSnapshotService,
)
from reportes.services_operating_finance_io import (
    OperatingFinanceExpenseImportService,
    OperatingFinanceTemplateService,
)
from reportes.services_budget_import import GeneralBudgetImportService
from reportes.services_budget_vs_actual import (
    BUDGET_VS_ACTUAL_SOURCE,
    BudgetCsvImportService,
    BudgetVsActualSnapshotService,
    write_example_budget_csv,
)
from reportes.services_budget_detail_import import (
    BudgetAuditMaterializationService,
    BudgetGeneralAuditService,
    TrustedBudgetDetailImportService,
)
from reportes.services_budget_area_upload import BudgetAreaUploadService
from reportes.services_budget_monitoring import BudgetMonitoringSnapshotService
from reportes.services_branch_admin_expense_import import BranchAdminExpenseImportService
from reportes.services_branch_real_operating_expense_import import (
    BranchRealOperatingExpenseImportService,
    BranchRealOperatingExpenseImportValidationError,
)
from reportes.services_historical_branch_expense_import import HistoricalBranchExpenseImportService
from reportes.services_operating_expense_automation import OperatingExpenseImportAutomationService
from reportes.services_production_expense_import import ProductionExpenseImportService


class OperatingFinanceBootstrapServiceTests(TestCase):
    def test_bootstrap_creates_catalog(self):
        Sucursal.objects.create(codigo="MAT", nombre="Matriz")

        summary = OperatingFinanceBootstrapService().bootstrap()

        self.assertGreaterEqual(summary["centros_costo"], 1)
        self.assertTrue(CentroCosto.objects.filter(codigo="CORP").exists())
        self.assertTrue(CategoriaGasto.objects.filter(codigo="MANO_OBRA_PROD").exists())
        self.assertTrue(CategoriaGasto.objects.filter(codigo="OPEX_TOTAL_SUC").exists())
        self.assertTrue(ReglaAsignacionGasto.objects.filter(categoria_gasto__codigo="RENTA_SUC").exists())


class ProductBusinessRuleValidationTests(TestCase):
    def test_collect_critical_product_business_rule_issues_reports_missing_rules(self):
        ProductBusinessRule.objects.all().delete()
        issues = collect_critical_product_business_rule_issues()

        self.assertEqual(len(issues), len(CRITICAL_FIXED_REVENTA_PRODUCT_NAMES))
        self.assertTrue(any("TE DEL JARDIN" in issue for issue in issues))

    def test_assert_critical_product_business_rules_present_raises_when_missing(self):
        ProductBusinessRule.objects.all().delete()
        with self.assertRaises(ImproperlyConfigured):
            assert_critical_product_business_rules_present()

    @override_settings(PRODUCT_BUSINESS_RULES_ENFORCE_ON_STARTUP=True)
    def test_collect_critical_product_business_rule_issues_accepts_seeded_rules(self):
        ProductBusinessRule.objects.all().delete()
        for product_name in CRITICAL_FIXED_REVENTA_PRODUCT_NAMES:
            ProductBusinessRule.objects.update_or_create(
                product_name=product_name,
                defaults={
                    "classification": ProductBusinessRule.CLASSIFICATION_REVENTA,
                    "is_fixed": True,
                },
            )

        issues = collect_critical_product_business_rule_issues()

        self.assertEqual(issues, [])


class OperatingFinanceSnapshotServiceTests(TestCase):
    def setUp(self):
        self.sucursal = Sucursal.objects.create(codigo="MAT", nombre="Matriz")
        self.point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=self.sucursal)
        self.point_product = PointProduct.objects.create(external_id="P1", sku="PFCMINI", name="Pastel Fresas con Crema Mini")
        self.receta = Receta.objects.create(
            nombre="Pastel Fresas con Crema Mini",
            codigo_point="PFCMINI",
            hash_contenido="hash_operating_finance",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        RecetaCostoSemanal.objects.create(
            scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
            identity_key=f"RECIPE:{self.receta.id}",
            label=self.receta.nombre,
            week_start=date(2026, 3, 23),
            week_end=date(2026, 3, 29),
            receta=self.receta,
            temporalidad=self.receta.temporalidad,
            temporalidad_detalle=self.receta.temporalidad_detalle,
            familia=self.receta.familia,
            categoria=self.receta.categoria,
            costo_mp=Decimal("58.806218"),
            costo_total=Decimal("58.806218"),
            metadata={},
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product,
            receta=self.receta,
            sale_date=date(2026, 3, 15),
            quantity=Decimal("10"),
            total_amount=Decimal("1850"),
            gross_amount=Decimal("1850"),
            net_amount=Decimal("1594.83"),
        )
        OperatingFinanceBootstrapService().bootstrap()
        self.prod_center = CentroCosto.objects.get(codigo="PROD")
        self.branch_center = CentroCosto.objects.get(codigo=f"SUC_{self.sucursal.codigo}")
        GastoOperativoMensual.objects.create(
            periodo=date(2026, 3, 1),
            centro_costo=self.prod_center,
            categoria_gasto=CategoriaGasto.objects.get(codigo="MANO_OBRA_PROD"),
            monto=Decimal("1000"),
        )
        GastoOperativoMensual.objects.create(
            periodo=date(2026, 3, 1),
            centro_costo=self.branch_center,
            categoria_gasto=CategoriaGasto.objects.get(codigo="RENTA_SUC"),
            monto=Decimal("500"),
        )
        GastoOperativoMensual.objects.create(
            periodo=date(2026, 3, 1),
            centro_costo=CentroCosto.objects.get(codigo="CORP"),
            categoria_gasto=CategoriaGasto.objects.get(codigo="ADMIN_CORP"),
            monto=Decimal("200"),
        )

    def test_build_snapshot_creates_financial_layers(self):
        summary = OperatingFinanceSnapshotService().build_snapshot(period_start=date(2026, 3, 1))

        self.assertEqual(summary.product_cost_rows, 1)
        self.assertEqual(summary.branch_contribution_rows, 1)
        product_row = ProductoCostoOperativoMensual.objects.get(periodo=date(2026, 3, 1), receta=self.receta)
        self.assertEqual(product_row.unidades_base, Decimal("10"))
        self.assertEqual(product_row.costo_mp_unit, Decimal("58.806218"))
        self.assertEqual(product_row.mano_obra_prod_unit, Decimal("100"))
        branch_row = ProductoSucursalContribucionMensual.objects.get(
            periodo=date(2026, 3, 1),
            receta=self.receta,
            sucursal=self.sucursal,
        )
        self.assertEqual(branch_row.gasto_comercial_total, Decimal("500"))
        company_row = EmpresaResultadoMensual.objects.get(periodo=date(2026, 3, 1))
        self.assertEqual(company_row.costo_materia_prima_total, Decimal("588.06"))
        self.assertEqual(company_row.costo_reventa_total, Decimal("0.00"))
        self.assertEqual(company_row.mano_obra_prod_total, Decimal("1000.00"))
        self.assertEqual(company_row.indirecto_prod_total, Decimal("0.00"))
        self.assertEqual(company_row.margen_bruto_total, Decimal("1261.94"))
        self.assertEqual(company_row.gasto_corporativo_total, Decimal("200"))
        pricing_row = ProductoPricingDecisionMensual.objects.get(periodo=date(2026, 3, 1), receta=self.receta)
        self.assertTrue(pricing_row.accion_sugerida)

    def test_build_snapshot_uses_profitability_sales_when_available(self):
        profitability_totals = {
            "rows": 1,
            "ventas_netas": Decimal("1700.00"),
            "costo_materia_prima": Decimal("600.00"),
            "costo_reventa": Decimal("50.00"),
            "gasto_fijo": Decimal("300.00"),
        }

        with patch.object(
            OperatingFinanceSnapshotService,
            "_profitability_totals",
            return_value=profitability_totals,
        ):
            OperatingFinanceSnapshotService().build_snapshot(period_start=date(2026, 3, 1))

        company_row = EmpresaResultadoMensual.objects.get(periodo=date(2026, 3, 1))
        self.assertEqual(company_row.venta_total, Decimal("1700.00"))
        self.assertEqual(company_row.margen_bruto_total, Decimal("1050.00"))
        self.assertEqual(company_row.gasto_comercial_total, Decimal("300.00"))
        self.assertEqual(company_row.utilidad_operativa_total, Decimal("550.00"))
        self.assertEqual(company_row.metadata["financial_totals_source"], "RENTABILIDAD_SUCURSAL")
        self.assertEqual(company_row.metadata["venta_total_calculada"], "1850.00")

    def test_build_snapshot_keeps_resale_cost_separate_from_manufacturing(self):
        resale_product = PointProduct.objects.create(
            external_id="PTE_RESALE",
            sku="TE-RESALE",
            name="TE DEL JARDIN",
            category="Te",
        )
        ProductBusinessRule.objects.create(
            product_name="TE DEL JARDIN",
            classification=ProductBusinessRule.CLASSIFICATION_REVENTA,
            is_fixed=True,
        )
        ProductoReventaCostoHistoricoMensual.objects.create(
            periodo=date(2026, 3, 1),
            producto_point=resale_product,
            costo_promedio=Decimal("12.50"),
        )
        cost_only_product = PointProduct.objects.create(
            external_id="PCOST_RESALE",
            sku="COST-RESALE",
            name="Producto con costo adquisicion",
            category="Bebidas",
        )
        ProductoReventaCostoHistoricoMensual.objects.create(
            periodo=date(2026, 3, 1),
            producto_point=cost_only_product,
            costo_promedio=Decimal("8.00"),
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=resale_product,
            receta=None,
            sale_date=date(2026, 3, 16),
            quantity=Decimal("4"),
            total_amount=Decimal("120"),
            gross_amount=Decimal("120"),
            net_amount=Decimal("103.45"),
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=cost_only_product,
            receta=None,
            sale_date=date(2026, 3, 17),
            quantity=Decimal("3"),
            total_amount=Decimal("90"),
            gross_amount=Decimal("90"),
            net_amount=Decimal("77.59"),
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            total_quantity=Decimal("17"),
            gross_amount=Decimal("2060"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("2060"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("1775.86"),
        )

        OperatingFinanceSnapshotService().build_snapshot(period_start=date(2026, 3, 1))

        company_row = EmpresaResultadoMensual.objects.get(periodo=date(2026, 3, 1))
        self.assertEqual(company_row.costo_materia_prima_total, Decimal("588.06"))
        self.assertEqual(company_row.costo_reventa_total, Decimal("74.00"))
        self.assertEqual(company_row.costo_fabricacion_total, Decimal("1588.06"))
        self.assertEqual(company_row.margen_bruto_total, Decimal("1397.94"))
        self.assertEqual(company_row.metadata["venta_reventa_total"], "210.00")
        self.assertEqual(company_row.metadata["reventa_rows"], 2)

    def test_build_snapshot_ignores_budget_expenses_for_real_financial_result(self):
        GastoOperativoMensual.objects.create(
            periodo=date(2026, 3, 1),
            centro_costo=self.prod_center,
            categoria_gasto=CategoriaGasto.objects.get(codigo="INDIRECTO_PROD"),
            monto=Decimal("9999"),
            tipo_dato=GastoOperativoMensual.TIPO_DATO_PRESUPUESTO,
        )

        OperatingFinanceSnapshotService().build_snapshot(period_start=date(2026, 3, 1))

        company_row = EmpresaResultadoMensual.objects.get(periodo=date(2026, 3, 1))
        product_row = ProductoCostoOperativoMensual.objects.get(periodo=date(2026, 3, 1), receta=self.receta)
        self.assertEqual(product_row.indirecto_prod_unit, Decimal("0"))
        self.assertEqual(company_row.indirecto_prod_total, Decimal("0.00"))
        self.assertEqual(company_row.utilidad_operativa_total, Decimal("-438.06"))

    def test_build_snapshot_does_not_use_future_weekly_cost_for_historical_month(self):
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product,
            receta=self.receta,
            sale_date=date(2026, 1, 15),
            quantity=Decimal("5"),
            total_amount=Decimal("925"),
            gross_amount=Decimal("925"),
            net_amount=Decimal("797.41"),
        )

        OperatingFinanceSnapshotService().build_snapshot(period_start=date(2026, 1, 1))

        product_row = ProductoCostoOperativoMensual.objects.get(periodo=date(2026, 1, 1), receta=self.receta)
        self.assertEqual(product_row.costo_mp_unit, Decimal("0"))
        self.assertEqual(product_row.costo_fabricacion_unit, Decimal("0"))
        self.assertEqual(product_row.metadata["cost_source"], "")

    def test_build_snapshot_zeroes_untrusted_product_cost_above_two_times_asp(self):
        RecetaCostoHistoricoMensual.objects.create(
            periodo=date(2026, 3, 1),
            receta=self.receta,
            costo_total=Decimal("50000"),
            costo_por_unidad_rendimiento=None,
            lineas_costeadas=1,
            lineas_totales=1,
            coverage_pct=Decimal("100"),
            metadata={"source": "corrupt_fixture"},
        )

        OperatingFinanceSnapshotService().build_snapshot(period_start=date(2026, 3, 1))

        product_row = ProductoCostoOperativoMensual.objects.get(periodo=date(2026, 3, 1), receta=self.receta)
        self.assertEqual(product_row.costo_mp_unit, Decimal("0"))
        self.assertEqual(product_row.mano_obra_prod_unit, Decimal("0"))
        self.assertEqual(product_row.costo_fabricacion_unit, Decimal("0"))
        self.assertTrue(product_row.metadata["guardrail_applied"])
        self.assertEqual(product_row.metadata["guardrail_reason"], "COSTO_FABRICACION_UNIT_GT_2X_ASP")
        self.assertEqual(product_row.metadata["raw_costo_mp_unit"], "50000.000000")

    def test_branch_expense_uses_exact_center_rule_before_generic_rule(self):
        categoria = CategoriaGasto.objects.get(codigo="RENTA_SUC")
        generic_rule = ReglaAsignacionGasto.objects.get(categoria_gasto=categoria, centro_costo__isnull=True)
        generic_rule.base_reparto = ReglaAsignacionGasto.BASE_VENTAS
        generic_rule.save(update_fields=["base_reparto"])
        second_point_product = PointProduct.objects.create(
            external_id="P2",
            sku="P2",
            name="Pastel Segundo",
        )
        second_receta = Receta.objects.create(
            nombre="Pastel Segundo",
            codigo_point="P2",
            hash_contenido="hash_operating_finance_2",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        RecetaCostoSemanal.objects.create(
            scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
            identity_key=f"RECIPE:{second_receta.id}",
            label=second_receta.nombre,
            week_start=date(2026, 3, 23),
            week_end=date(2026, 3, 29),
            receta=second_receta,
            temporalidad=second_receta.temporalidad,
            temporalidad_detalle=second_receta.temporalidad_detalle,
            familia=second_receta.familia,
            categoria=second_receta.categoria,
            costo_mp=Decimal("40"),
            costo_total=Decimal("40"),
            metadata={},
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=second_point_product,
            receta=second_receta,
            sale_date=date(2026, 3, 16),
            quantity=Decimal("5"),
            total_amount=Decimal("1500"),
            gross_amount=Decimal("1500"),
            net_amount=Decimal("1293.10"),
        )
        ReglaAsignacionGasto.objects.create(
            nombre="Renta Matriz por unidades",
            categoria_gasto=categoria,
            centro_costo=self.branch_center,
            base_reparto=ReglaAsignacionGasto.BASE_UNIDADES,
            prioridad=1,
            activo=True,
        )

        summary = OperatingFinanceSnapshotService().build_snapshot(period_start=date(2026, 3, 1))

        self.assertEqual(summary.branch_contribution_rows, 2)
        branch_row = ProductoSucursalContribucionMensual.objects.get(
            periodo=date(2026, 3, 1),
            receta=self.receta,
            sucursal=self.sucursal,
        )
        # Con reparto por unidades: 10 de 15 piezas = 333.3333...
        self.assertEqual(branch_row.gasto_comercial_total.quantize(Decimal("0.01")), Decimal("333.33"))

    def test_company_sales_total_prefers_official_month_and_tracks_unmapped_gap(self):
        mapped_non_recipe = Receta.objects.create(
            nombre="Letrero Chispas Felicidades",
            codigo_point="0179",
            hash_contenido="hash_non_recipe_mapped",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Pillines",
            sheet_name="AUTO_POINT_SALES",
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=PointProduct.objects.create(external_id="PX", sku="PX", name="Venta sin receta"),
            receta=None,
            sale_date=date(2026, 3, 18),
            quantity=Decimal("2"),
            total_amount=Decimal("400"),
            gross_amount=Decimal("400"),
            net_amount=Decimal("344.83"),
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=PointProduct.objects.create(external_id="PVELA", sku="VELA1", name="Vela Bengala", category="Granmark"),
            receta=None,
            sale_date=date(2026, 3, 19),
            quantity=Decimal("1"),
            total_amount=Decimal("250"),
            gross_amount=Decimal("250"),
            net_amount=Decimal("215.52"),
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=PointProduct.objects.create(
                external_id="PLET",
                sku="0179",
                name="Letrero Chispas Felicidades",
                category="Pillines",
            ),
            receta=mapped_non_recipe,
            sale_date=date(2026, 3, 20),
            quantity=Decimal("1"),
            total_amount=Decimal("170"),
            gross_amount=Decimal("170"),
            net_amount=Decimal("146.55"),
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            total_quantity=Decimal("14"),
            gross_amount=Decimal("2670"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("2670"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("2301.73"),
        )

        OperatingFinanceSnapshotService().build_snapshot(period_start=date(2026, 3, 1))

        company_row = EmpresaResultadoMensual.objects.get(periodo=date(2026, 3, 1))
        self.assertEqual(company_row.venta_total, Decimal("2670"))
        self.assertEqual(company_row.metadata["sales_total_source"], "POINT_MONTHLY_OFFICIAL")
        self.assertEqual(company_row.metadata["venta_costeada_total"], "1850.00")
        self.assertEqual(company_row.metadata["venta_sin_mapear_total"], "820.00")
        self.assertEqual(company_row.metadata["venta_no_receta_total"], "420.00")
        self.assertEqual(company_row.metadata["venta_receta_sin_match_total"], "400.00")

    def test_company_sales_total_uses_canonical_range_source_when_monthly_official_missing(self):
        PointDailySale.objects.filter().delete()
        PointMonthlySalesOfficial.objects.filter().delete()
        point_product_v2 = PointProduct.objects.create(external_id="P2", sku="PFCMINI", name="Pastel Fresas con Crema Mini")
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=point_product_v2,
            receta=self.receta,
            sale_date=date(2026, 3, 16),
            quantity=Decimal("50"),
            total_amount=Decimal("5000"),
            gross_amount=Decimal("5000"),
            net_amount=Decimal("4310.34"),
            source_endpoint="/Report/VentasCategorias",
        )
        from pos_bridge.models import PointSalesDailyCategoryFact

        PointSalesDailyCategoryFact.objects.create(
            branch=self.point_branch,
            sale_date=date(2026, 3, 16),
            sucursal_nombre=self.sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("10"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("1850"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("1850"),
        )

        OperatingFinanceSnapshotService().build_snapshot(period_start=date(2026, 3, 1))

        company_row = EmpresaResultadoMensual.objects.get(periodo=date(2026, 3, 1))
        self.assertEqual(company_row.venta_total, Decimal("1850"))
        self.assertEqual(company_row.metadata["sales_total_source"], "SALES_READ_V2_FACT")

    def test_audit_operating_finance_sources_reports_classification_buckets_and_balance(self):
        service_recipe = Receta.objects.create(
            nombre="Servicio Decoracion Dashboard",
            codigo_point="SERV001",
            hash_contenido="hash_audit_service",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            modo_costeo=Receta.MODO_COSTEO_SERVICIO,
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=PointProduct.objects.create(
                external_id="PSERV",
                sku="SERV001",
                name="Servicio Decoracion Dashboard",
                category="Pasteles",
            ),
            receta=service_recipe,
            sale_date=date(2026, 3, 18),
            quantity=Decimal("1"),
            total_amount=Decimal("120"),
            gross_amount=Decimal("120"),
            net_amount=Decimal("103.45"),
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=PointProduct.objects.create(
                external_id="PCAND",
                sku="CAND001",
                name="Producto Pendiente Match",
                category="Pasteles",
            ),
            receta=None,
            sale_date=date(2026, 3, 19),
            quantity=Decimal("2"),
            total_amount=Decimal("80"),
            gross_amount=Decimal("80"),
            net_amount=Decimal("68.97"),
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            total_quantity=Decimal("13"),
            gross_amount=Decimal("1300"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("1300"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("1120.69"),
        )

        stdout = StringIO()
        call_command("audit_operating_finance_sources", period="2026-03", stdout=stdout)
        payload = json.loads(stdout.getvalue())

        self.assertEqual(payload["sources"]["daily_sales_total"], "2050.00")
        self.assertEqual(payload["sources"]["canonical_range_total"], "2050.00")
        self.assertEqual(payload["sources"]["official_month_total"], "1300.00")
        self.assertEqual(payload["commercial_classification"]["venta_costeada_total"], "1850.00")
        self.assertEqual(payload["commercial_classification"]["venta_no_receta_total"], "120.00")
        self.assertEqual(payload["commercial_classification"]["venta_receta_sin_match_total"], "80.00")
        self.assertEqual(payload["commercial_classification"]["venta_sin_mapear_total"], "200.00")
        self.assertEqual(payload["commercial_classification"]["non_recipe_bucket_totals"]["SERVICIO"], "120.00")
        self.assertEqual(payload["commercial_classification"]["non_recipe_bucket_counts"]["SERVICIO"], 1)
        self.assertEqual(payload["commercial_classification"]["row_counts"]["with_recipe_id"], 2)
        self.assertEqual(payload["commercial_classification"]["row_counts"]["without_recipe_id"], 1)
        self.assertTrue(payload["validations"]["classification_balance"]["balanced"])
        self.assertEqual(payload["commercial_classification"]["top_non_recipe_products"][0]["bucket"], "SERVICIO")
        self.assertEqual(payload["commercial_classification"]["top_candidate_recipe_rows"][0]["product_name"], "Producto Pendiente Match")


class MonthlyHistoricalCostingServiceTests(TestCase):
    def setUp(self):
        self.unit_kg = UnidadMedida.objects.create(codigo="kg", nombre="Kilogramo", tipo=UnidadMedida.TIPO_MASA, factor_to_base=1000)
        self.unit_g = UnidadMedida.objects.create(codigo="g", nombre="Gramo", tipo=UnidadMedida.TIPO_MASA, factor_to_base=1)
        self.unit_lt = UnidadMedida.objects.create(codigo="lt", nombre="Litro", tipo=UnidadMedida.TIPO_VOLUMEN, factor_to_base=1000)
        self.unit_ml = UnidadMedida.objects.create(codigo="ml", nombre="Mililitro", tipo=UnidadMedida.TIPO_VOLUMEN, factor_to_base=1)
        self.unit_pza = UnidadMedida.objects.create(codigo="pza", nombre="Pieza", tipo=UnidadMedida.TIPO_PIEZA, factor_to_base=1)
        self.harvest = Insumo.objects.create(
            codigo="HARINA",
            nombre="Harina",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unit_kg,
            activo=True,
        )
        self.sugar = Insumo.objects.create(
            codigo="AZUCAR",
            nombre="Azucar",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unit_kg,
            activo=True,
        )
        self.prep_recipe = Receta.objects.create(
            nombre="Batido Base",
            hash_contenido="hist_prep_hash",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("2"),
            rendimiento_unidad=self.unit_kg,
        )
        self.prep_insumo = Insumo.objects.create(
            codigo=f"DERIVADO:RECETA:{self.prep_recipe.id}:PREPARACION",
            codigo_point="PREPBASE",
            nombre="Batido Base",
            tipo_item=Insumo.TIPO_INTERNO,
            unidad_base=self.unit_kg,
            activo=True,
        )
        LineaReceta.objects.create(
            receta=self.prep_recipe,
            posicion=1,
            insumo=self.harvest,
            insumo_texto="Harina",
            cantidad=Decimal("1"),
            unidad=self.unit_kg,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("0"),
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
        )
        LineaReceta.objects.create(
            receta=self.prep_recipe,
            posicion=2,
            insumo=self.sugar,
            insumo_texto="Azucar",
            cantidad=Decimal("1"),
            unidad=self.unit_kg,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("0"),
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
        )
        self.final_recipe = Receta.objects.create(
            nombre="Pastel Historico",
            codigo_point="PHIST",
            hash_contenido="hist_final_hash",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        LineaReceta.objects.create(
            receta=self.final_recipe,
            posicion=1,
            insumo=self.prep_insumo,
            insumo_texto="Batido Base",
            cantidad=Decimal("0.5"),
            unidad=self.unit_kg,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("0"),
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
        )
        CostoInsumo.objects.create(
            insumo=self.harvest,
            fecha=date(2026, 1, 5),
            costo_unitario=Decimal("10"),
            source_hash="harina_jan_1",
            raw={"fecha": "2026-01-05", "precio": 100, "cantidad": 10, "producto": "Harina"},
        )
        CostoInsumo.objects.create(
            insumo=self.harvest,
            fecha=date(2026, 1, 20),
            costo_unitario=Decimal("20"),
            source_hash="harina_jan_2",
            raw={"fecha": "2026-01-20", "precio": 200, "cantidad": 10, "producto": "Harina"},
        )
        CostoInsumo.objects.create(
            insumo=self.sugar,
            fecha=date(2025, 12, 20),
            costo_unitario=Decimal("30"),
            source_hash="azucar_dec",
            raw={"fecha": "2025-12-20", "precio": 300, "cantidad": 10, "producto": "Azucar"},
        )

    def test_build_period_creates_weighted_and_rolled_forward_costs(self):
        Sucursal.objects.create(codigo="MAT", nombre="Matriz")
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz")
        point_product = PointProduct.objects.create(external_id="PX1", sku="PHIST", name="Pastel Historico")
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=self.final_recipe,
            sale_date=date(2026, 1, 15),
            quantity=Decimal("3"),
            total_amount=Decimal("300"),
            gross_amount=Decimal("300"),
            net_amount=Decimal("258.62"),
        )

        summary = MonthlyHistoricalCostingService().build_period(period_start=date(2026, 1, 1))

        self.assertGreaterEqual(summary.insumo_rows, 2)
        harina = InsumoCostoHistoricoMensual.objects.get(periodo=date(2026, 1, 1), insumo=self.harvest)
        azucar = InsumoCostoHistoricoMensual.objects.get(periodo=date(2026, 1, 1), insumo=self.sugar)
        self.assertEqual(harina.costo_unitario, Decimal("15.000000"))
        self.assertEqual(harina.metodo, InsumoCostoHistoricoMensual.METODO_PROMEDIO_MENSUAL)
        self.assertEqual(azucar.costo_unitario, Decimal("30.000000"))
        self.assertEqual(azucar.metodo, InsumoCostoHistoricoMensual.METODO_ARRASTRE)

        prep_cost = RecetaCostoHistoricoMensual.objects.get(periodo=date(2026, 1, 1), receta=self.prep_recipe)
        final_cost = RecetaCostoHistoricoMensual.objects.get(periodo=date(2026, 1, 1), receta=self.final_recipe)
        self.assertEqual(prep_cost.costo_total, Decimal("45.000000"))
        self.assertEqual(prep_cost.costo_por_unidad_rendimiento, Decimal("22.500000"))
        self.assertEqual(final_cost.costo_total, Decimal("11.250000"))
        self.assertEqual(final_cost.coverage_pct, Decimal("100.000000"))

    def test_monthly_cost_normalizes_point_kg_cost_to_gram_base(self):
        chocolate = Insumo.objects.create(
            codigo="CHOCOLATE",
            nombre="Chocolate",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unit_g,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=chocolate,
            fecha=date(2026, 4, 10),
            costo_unitario=Decimal("319.83"),
            source_hash="chocolate_apr_kg",
            raw={"fecha": "2026-04-10", "cantidad": 1, "unit": "kg", "producto": "Chocolate"},
        )

        row = MonthlyHistoricalCostingService()._build_insumo_monthly_cost(
            period_start=date(2026, 4, 1),
            insumo=chocolate,
        )

        self.assertIsNotNone(row)
        self.assertEqual(row.costo_unitario, Decimal("0.319830"))
        self.assertEqual(row.metadata["unit_normalization"][0]["source_unit"], "kg")
        self.assertEqual(row.metadata["unit_normalization"][0]["target_unit"], "g")

    def test_monthly_cost_normalizes_point_liter_cost_to_ml_base(self):
        vanilla = Insumo.objects.create(
            codigo="VAINILLA",
            nombre="Vainilla",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unit_ml,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=vanilla,
            fecha=date(2026, 4, 11),
            costo_unitario=Decimal("69.62175"),
            source_hash="vainilla_apr_lt",
            raw={"fecha": "2026-04-11", "cantidad": 1, "unidad": "litro", "producto": "Vainilla"},
        )

        row = MonthlyHistoricalCostingService()._build_insumo_monthly_cost(
            period_start=date(2026, 4, 1),
            insumo=vanilla,
        )

        self.assertIsNotNone(row)
        self.assertEqual(row.costo_unitario, Decimal("0.069622"))
        self.assertEqual(row.metadata["unit_normalization"][0]["source_unit"], "lt")
        self.assertEqual(row.metadata["unit_normalization"][0]["target_unit"], "ml")

    def test_recipe_historical_cost_uses_normalized_insumo_base_cost(self):
        flour = Insumo.objects.create(
            codigo="HARINA_G",
            nombre="Harina por gramo",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unit_g,
            activo=True,
        )
        recipe = Receta.objects.create(
            nombre="Base Gramos",
            hash_contenido="hist_grams_hash",
            tipo=Receta.TIPO_PREPARACION,
            rendimiento_cantidad=Decimal("1"),
            rendimiento_unidad=self.unit_kg,
        )
        LineaReceta.objects.create(
            receta=recipe,
            posicion=1,
            insumo=flour,
            insumo_texto="Harina por gramo",
            cantidad=Decimal("500"),
            unidad=self.unit_g,
            unidad_texto="g",
            costo_unitario_snapshot=Decimal("0"),
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
        )
        CostoInsumo.objects.create(
            insumo=flour,
            fecha=date(2026, 4, 12),
            costo_unitario=Decimal("18.50"),
            source_hash="harina_apr_kg",
            raw={"fecha": "2026-04-12", "cantidad": 1, "unit": "kg", "producto": "Harina"},
        )

        result = MonthlyHistoricalCostingService()._build_recipe_monthly_cost(
            period_start=date(2026, 4, 1),
            receta=recipe,
        )

        self.assertEqual(result.total_cost, Decimal("9.250000"))
        self.assertEqual(result.unit_cost, Decimal("9.250000"))

    def test_monthly_cost_infers_missing_raw_unit_from_same_insumo_point_cost(self):
        crunch = Insumo.objects.create(
            codigo="CRUNCH",
            nombre="Trozos Crunch",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unit_g,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=crunch,
            fecha=date(2026, 2, 28),
            costo_unitario=Decimal("264.021944"),
            source_hash="crunch_feb_no_unit",
            raw={"fuente": "AUTO_HOMOLOGACION_TOKEN_SET", "match_nombre": "Crunch Trozos"},
        )
        CostoInsumo.objects.create(
            insumo=crunch,
            fecha=date(2026, 4, 20),
            costo_unitario=Decimal("264.022200"),
            source_hash="crunch_apr_point_kg",
            raw={"source": "POINT_EXISTENCIA_ALMACEN", "unit": "KG", "point_name": "TROZOS CRUNCH"},
        )

        row = MonthlyHistoricalCostingService()._build_insumo_monthly_cost(
            period_start=date(2026, 2, 1),
            insumo=crunch,
        )

        self.assertIsNotNone(row)
        self.assertEqual(row.costo_unitario, Decimal("0.264022"))
        normalization = row.metadata["unit_normalization"][0]
        self.assertEqual(normalization["unit_resolution"], "inferred_from_same_insumo_cost")
        self.assertEqual(normalization["inferred_from_unit"], "kg")

    def test_operating_snapshot_uses_monthly_historical_cost_when_weekly_missing(self):
        sucursal = Sucursal.objects.create(codigo="MAT", nombre="Matriz")
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=sucursal)
        point_product = PointProduct.objects.create(external_id="PX2", sku="PHIST", name="Pastel Historico")
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=self.final_recipe,
            sale_date=date(2026, 1, 15),
            quantity=Decimal("4"),
            total_amount=Decimal("400"),
            gross_amount=Decimal("400"),
            net_amount=Decimal("344.83"),
        )
        MonthlyHistoricalCostingService().build_period(period_start=date(2026, 1, 1))
        OperatingFinanceBootstrapService().bootstrap()
        GastoOperativoMensual.objects.create(
            periodo=date(2026, 1, 1),
            centro_costo=CentroCosto.objects.get(codigo="PROD"),
            categoria_gasto=CategoriaGasto.objects.get(codigo="MANO_OBRA_PROD"),
            monto=Decimal("100"),
        )

        summary = OperatingFinanceSnapshotService().build_snapshot(period_start=date(2026, 1, 1))

        self.assertEqual(summary.product_cost_rows, 1)
        row = ProductoCostoOperativoMensual.objects.get(periodo=date(2026, 1, 1), receta=self.final_recipe)
        self.assertEqual(row.costo_mp_unit, Decimal("11.250000"))

    def test_alias_rule_uses_reference_insumo_cost(self):
        fecula = Insumo.objects.create(
            codigo="MAICENA",
            nombre="Maicena",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unit_kg,
            activo=True,
        )
        almidon = Insumo.objects.create(
            codigo="ALMIDON",
            nombre="Almidón",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unit_kg,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=almidon,
            fecha=date(2026, 1, 18),
            costo_unitario=Decimal("21.44"),
            source_hash="almidon_jan_1",
            raw={"fecha": "2026-01-18", "precio": 214.4, "cantidad": 10, "producto": "Almidón"},
        )
        ReglaCostoHistoricoInsumo.objects.create(
            insumo_origen=fecula,
            metodo=ReglaCostoHistoricoInsumo.METODO_EQUIVALENCIA,
            insumo_referencia=almidon,
            prioridad=10,
            activo=True,
        )

        service = MonthlyHistoricalCostingService()
        row = service._build_insumo_monthly_cost(period_start=date(2026, 1, 1), insumo=fecula)
        self.assertIsNotNone(row)
        row = InsumoCostoHistoricoMensual.objects.get(periodo=date(2026, 1, 1), insumo=fecula)
        self.assertEqual(row.costo_unitario, Decimal("21.440000"))
        self.assertEqual(row.metodo, InsumoCostoHistoricoMensual.METODO_EQUIVALENCIA)

    def test_next_known_rule_uses_first_future_cost(self):
        cream = Insumo.objects.create(
            codigo="CREMA",
            nombre="Sustituto de crema",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unit_kg,
            activo=True,
        )
        ReglaCostoHistoricoInsumo.objects.create(
            insumo_origen=cream,
            metodo=ReglaCostoHistoricoInsumo.METODO_SIGUIENTE,
            prioridad=20,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=cream,
            fecha=date(2026, 3, 21),
            costo_unitario=Decimal("80"),
            source_hash="cream_mar_1",
            raw={"fecha": "2026-03-21", "precio": 800, "cantidad": 10, "producto": "Sustituto de crema"},
        )

        service = MonthlyHistoricalCostingService()
        row = service._build_insumo_monthly_cost(period_start=date(2026, 2, 1), insumo=cream)
        self.assertIsNotNone(row)
        row = InsumoCostoHistoricoMensual.objects.get(periodo=date(2026, 2, 1), insumo=cream)
        self.assertEqual(row.costo_unitario, Decimal("80.000000"))
        self.assertEqual(row.metodo, InsumoCostoHistoricoMensual.METODO_SIGUIENTE)

    def test_derived_presentation_uses_parent_unit_cost_in_historical_snapshot(self):
        parent_recipe = Receta.objects.create(
            nombre="Pay Grande",
            codigo_point="PAYG",
            hash_contenido="hist_parent_hash",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            rendimiento_cantidad=Decimal("1"),
            rendimiento_unidad=self.unit_pza,
        )
        LineaReceta.objects.create(
            receta=parent_recipe,
            posicion=1,
            insumo=self.harvest,
            insumo_texto="Harina",
            cantidad=Decimal("2"),
            unidad=self.unit_kg,
            unidad_texto="kg",
            costo_unitario_snapshot=Decimal("0"),
            match_status=LineaReceta.STATUS_AUTO,
            match_method=LineaReceta.MATCH_EXACT,
        )
        derived_recipe = Receta.objects.create(
            nombre="Pay Grande Rebanada",
            codigo_point="PAYR",
            hash_contenido="hist_derived_hash",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            rendimiento_cantidad=Decimal("1"),
            rendimiento_unidad=self.unit_pza,
        )
        RecetaPresentacionDerivada.objects.create(
            receta_padre=parent_recipe,
            receta_derivada=derived_recipe,
            codigo_point_derivado="PAYR",
            nombre_derivado="Pay Grande Rebanada",
            unidades_por_padre=Decimal("8"),
            requiere_componentes_directos=False,
        )
        point_branch = PointBranch.objects.create(external_id="9", name="Sucursal 9")
        point_product = PointProduct.objects.create(external_id="PX9", sku="PAYR", name="Pay Grande Rebanada")
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=derived_recipe,
            sale_date=date(2026, 1, 10),
            quantity=Decimal("4"),
            total_amount=Decimal("400"),
            gross_amount=Decimal("400"),
            net_amount=Decimal("344.83"),
        )

        MonthlyHistoricalCostingService().build_period(period_start=date(2026, 1, 1))

        parent_cost = RecetaCostoHistoricoMensual.objects.get(periodo=date(2026, 1, 1), receta=parent_recipe)
        derived_cost = RecetaCostoHistoricoMensual.objects.get(periodo=date(2026, 1, 1), receta=derived_recipe)
        self.assertEqual(parent_cost.costo_total, Decimal("30.000000"))
        self.assertEqual(derived_cost.costo_total, Decimal("3.750000"))
        self.assertEqual(derived_cost.coverage_pct, Decimal("100.000000"))

    def test_historical_snapshot_excludes_service_accessory_sales(self):
        service_recipe = Receta.objects.create(
            nombre="Letrero Chispas Felicidades",
            codigo_point="0179",
            hash_contenido="hist_service_hash",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            modo_costeo=Receta.MODO_COSTEO_SERVICIO,
        )
        point_branch = PointBranch.objects.create(external_id="10", name="Sucursal 10")
        point_product = PointProduct.objects.create(external_id="PSV", sku="0179", name="Letrero Chispas Felicidades", category="Pillines")
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=service_recipe,
            sale_date=date(2026, 1, 12),
            quantity=Decimal("2"),
            total_amount=Decimal("100"),
            gross_amount=Decimal("100"),
            net_amount=Decimal("86.21"),
        )

        MonthlyHistoricalCostingService().build_period(period_start=date(2026, 1, 1))

        self.assertFalse(
            RecetaCostoHistoricoMensual.objects.filter(periodo=date(2026, 1, 1), receta=service_recipe).exists()
        )

    def test_historical_snapshot_includes_resale_sales_using_purchase_cost(self):
        tea_insumo = Insumo.objects.create(
            codigo="TE01",
            codigo_point="TE01",
            nombre="Te Chai",
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            unidad_base=self.unit_pza,
            activo=True,
        )
        CostoInsumo.objects.create(
            insumo=tea_insumo,
            fecha=date(2026, 1, 5),
            costo_unitario=Decimal("12.50"),
            source_hash="hist_resale_cost_te01",
            raw={"source": "POINT_COMPRAS_HISTORICAS", "cantidad": "10"},
        )
        resale_recipe = Receta.objects.create(
            nombre="Te Chai",
            codigo_point="TE01",
            hash_contenido="hist_resale_hash",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            modo_costeo=Receta.MODO_COSTEO_REVENTA,
        )
        point_branch = PointBranch.objects.create(external_id="11", name="Sucursal 11")
        point_product = PointProduct.objects.create(external_id="PTE", sku="TE01", name="Te Chai", category="Te")
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=resale_recipe,
            sale_date=date(2026, 1, 12),
            quantity=Decimal("2"),
            total_amount=Decimal("100"),
            gross_amount=Decimal("100"),
            net_amount=Decimal("86.21"),
        )

        MonthlyHistoricalCostingService().build_period(period_start=date(2026, 1, 1))

        historical_row = RecetaCostoHistoricoMensual.objects.get(periodo=date(2026, 1, 1), receta=resale_recipe)
        self.assertEqual(historical_row.costo_total, Decimal("12.500000"))
        self.assertEqual(historical_row.coverage_pct, Decimal("100.000000"))

    def test_historical_snapshot_includes_resale_sales_using_alias_match(self):
        box_insumo = Insumo.objects.create(
            codigo="042",
            codigo_point="042",
            nombre="CAJA G",
            tipo_item=Insumo.TIPO_EMPAQUE,
            unidad_base=self.unit_pza,
            activo=True,
        )
        InsumoAlias.objects.create(nombre="CAJA G PARA VENTA", insumo=box_insumo)
        CostoInsumo.objects.create(
            insumo=box_insumo,
            fecha=date(2026, 3, 2),
            costo_unitario=Decimal("10.47"),
            source_hash="hist_resale_cost_boxg",
            raw={"source": "POINT_COMPRAS_HISTORICAS", "cantidad": "10"},
        )
        ReglaCostoHistoricoInsumo.objects.create(
            insumo_origen=box_insumo,
            metodo=ReglaCostoHistoricoInsumo.METODO_SIGUIENTE,
            prioridad=20,
            activo=True,
            notas="Usar primer costo posterior disponible para caja vendida directo.",
        )
        resale_recipe = Receta.objects.create(
            nombre="CAJA G PARA VENTA",
            codigo_point="0239",
            hash_contenido="hist_resale_box_hash",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            modo_costeo=Receta.MODO_COSTEO_REVENTA,
        )
        point_branch = PointBranch.objects.create(external_id="12", name="Sucursal 12")
        point_product = PointProduct.objects.create(external_id="PBOX", sku="0239", name="CAJA G PARA VENTA", category="Industrias lec")
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=resale_recipe,
            sale_date=date(2026, 1, 12),
            quantity=Decimal("2"),
            total_amount=Decimal("40"),
            gross_amount=Decimal("40"),
            net_amount=Decimal("34.48"),
        )

        MonthlyHistoricalCostingService().build_period(period_start=date(2026, 1, 1))

        historical_row = RecetaCostoHistoricoMensual.objects.get(periodo=date(2026, 1, 1), receta=resale_recipe)
        self.assertEqual(historical_row.costo_total, Decimal("10.470000"))
        self.assertEqual(historical_row.coverage_pct, Decimal("100.000000"))

    def test_historical_snapshot_freezes_non_recipe_resale_product_cost(self):
        point_branch = PointBranch.objects.create(external_id="13", name="Sucursal 13")
        point_product = PointProduct.objects.create(
            external_id="COCA450",
            sku="COCA450",
            name="COCA-COLA 450 ML",
            category="Bebidas",
        )
        ProductoReventaCosto.objects.create(
            producto_point=point_product,
            costo_unitario=Decimal("13.25"),
            fecha_vigencia=date(2026, 3, 15),
            fuente=ProductoReventaCosto.FUENTE_POINT_ALMACEN,
            unidad="pza",
            cantidad_snapshot=Decimal("24"),
            source_hash="hist_resale_product_coca",
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=None,
            sale_date=date(2026, 3, 18),
            quantity=Decimal("2"),
            total_amount=Decimal("50"),
            gross_amount=Decimal("50"),
            net_amount=Decimal("50"),
        )

        summary = MonthlyHistoricalCostingService().build_period(period_start=date(2026, 3, 1))

        historical_row = ProductoReventaCostoHistoricoMensual.objects.get(
            periodo=date(2026, 3, 1),
            producto_point=point_product,
        )
        self.assertEqual(summary.producto_reventa_rows, 1)
        self.assertEqual(historical_row.costo_promedio, Decimal("13.250000"))
        self.assertEqual(historical_row.metodo, ProductoReventaCostoHistoricoMensual.METODO_POINT_ALMACEN)


class OperatingFinanceResaleTests(TestCase):
    FORCED_RESALE_CASES = [
        {"name": "CAJA CH PARA VENTA", "sku": "0240", "category": "Industrias lec", "amount": Decimal("30"), "recipe_mode": None},
        {"name": "CAJA G PARA VENTA", "sku": "0239", "category": "Industrias lec", "amount": Decimal("80"), "recipe_mode": Receta.MODO_COSTEO_REVENTA},
        {"name": "COCA-COLA 450 ML", "sku": "COCA450", "category": "Coca-cola", "amount": Decimal("66"), "recipe_mode": None},
        {"name": "TE DEL JARDIN", "sku": "0313", "category": "TE", "amount": Decimal("120"), "recipe_mode": Receta.MODO_COSTEO_REVENTA},
        {"name": "CAFE STARBUCKS FRAPPUCINO", "sku": "84111506", "category": "TE", "amount": Decimal("140"), "recipe_mode": Receta.MODO_COSTEO_REVENTA},
    ]

    def test_mapped_resale_sale_counts_as_costed_not_non_recipe(self):
        sucursal = Sucursal.objects.create(codigo="MAT", nombre="Matriz")
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=sucursal)
        point_product = PointProduct.objects.create(external_id="PTE", sku="TE01", name="Te Chai", category="Te")
        receta = Receta.objects.create(
            nombre="Te Chai",
            codigo_point="TE01",
            hash_contenido="hash_operating_finance_resale",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            modo_costeo=Receta.MODO_COSTEO_REVENTA,
        )
        RecetaCostoSemanal.objects.create(
            scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
            identity_key=f"RECIPE:{receta.id}",
            label=receta.nombre,
            week_start=date(2026, 3, 23),
            week_end=date(2026, 3, 29),
            receta=receta,
            temporalidad=receta.temporalidad,
            temporalidad_detalle=receta.temporalidad_detalle,
            familia=receta.familia,
            categoria=receta.categoria,
            costo_mp=Decimal("12.50"),
            costo_total=Decimal("12.50"),
            metadata={},
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=receta,
            sale_date=date(2026, 3, 15),
            quantity=Decimal("10"),
            total_amount=Decimal("200"),
            gross_amount=Decimal("200"),
            net_amount=Decimal("172.41"),
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            total_quantity=Decimal("10"),
            gross_amount=Decimal("200"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("200"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("172.41"),
        )
        OperatingFinanceBootstrapService().bootstrap()

        OperatingFinanceSnapshotService().build_snapshot(period_start=date(2026, 3, 1))

        company_row = EmpresaResultadoMensual.objects.get(periodo=date(2026, 3, 1))
        self.assertEqual(company_row.metadata["venta_costeada_total"], "200.00")
        self.assertEqual(company_row.metadata["venta_no_receta_total"], "0")
        self.assertEqual(company_row.metadata["venta_receta_sin_match_total"], "0")

    def test_resolved_resale_sale_without_receta_id_counts_as_costed_not_candidate(self):
        sucursal = Sucursal.objects.create(codigo="MAT", nombre="Matriz")
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=sucursal)
        point_product = PointProduct.objects.create(external_id="PTE2", sku="TE02", name="Te Matcha", category="Te")
        receta = Receta.objects.create(
            nombre="Te Matcha",
            codigo_point="TE02",
            hash_contenido="hash_operating_finance_resale_resolved",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            modo_costeo=Receta.MODO_COSTEO_REVENTA,
        )
        RecetaCostoSemanal.objects.create(
            scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
            identity_key=f"RECIPE:{receta.id}",
            label=receta.nombre,
            week_start=date(2026, 3, 23),
            week_end=date(2026, 3, 29),
            receta=receta,
            temporalidad=receta.temporalidad,
            temporalidad_detalle=receta.temporalidad_detalle,
            familia=receta.familia,
            categoria=receta.categoria,
            costo_mp=Decimal("8.50"),
            costo_total=Decimal("8.50"),
            metadata={},
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=None,
            sale_date=date(2026, 3, 15),
            quantity=Decimal("10"),
            total_amount=Decimal("200"),
            gross_amount=Decimal("200"),
            net_amount=Decimal("172.41"),
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            total_quantity=Decimal("10"),
            gross_amount=Decimal("200"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("200"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("172.41"),
        )
        OperatingFinanceBootstrapService().bootstrap()

        OperatingFinanceSnapshotService().build_snapshot(period_start=date(2026, 3, 1))

        company_row = EmpresaResultadoMensual.objects.get(periodo=date(2026, 3, 1))
        contribution_row = ProductoSucursalContribucionMensual.objects.get(periodo=date(2026, 3, 1), receta=receta, sucursal=sucursal)
        self.assertEqual(contribution_row.venta_total, Decimal("200"))
        self.assertEqual(company_row.metadata["venta_costeada_total"], "200.00")
        self.assertEqual(company_row.metadata["venta_no_receta_total"], "0")
        self.assertEqual(company_row.metadata["venta_receta_sin_match_total"], "0")

    def test_audit_excludes_resolved_resale_sale_without_receta_id_from_candidate_total(self):
        sucursal = Sucursal.objects.create(codigo="MAT", nombre="Matriz")
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=sucursal)
        resolvable_product = PointProduct.objects.create(external_id="PTE3", sku="TE03", name="Te Jazmin", category="Te")
        unresolved_product = PointProduct.objects.create(external_id="PUNK", sku="UNK01", name="Producto Pendiente Match", category="Pasteles")
        receta = Receta.objects.create(
            nombre="Te Jazmin",
            codigo_point="TE03",
            hash_contenido="hash_operating_finance_resale_audit",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            modo_costeo=Receta.MODO_COSTEO_REVENTA,
        )
        RecetaCostoSemanal.objects.create(
            scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
            identity_key=f"RECIPE:{receta.id}",
            label=receta.nombre,
            week_start=date(2026, 3, 23),
            week_end=date(2026, 3, 29),
            receta=receta,
            temporalidad=receta.temporalidad,
            temporalidad_detalle=receta.temporalidad_detalle,
            familia=receta.familia,
            categoria=receta.categoria,
            costo_mp=Decimal("6.00"),
            costo_total=Decimal("6.00"),
            metadata={},
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=resolvable_product,
            receta=None,
            sale_date=date(2026, 3, 18),
            quantity=Decimal("5"),
            total_amount=Decimal("100"),
            gross_amount=Decimal("100"),
            net_amount=Decimal("86.21"),
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=unresolved_product,
            receta=None,
            sale_date=date(2026, 3, 19),
            quantity=Decimal("2"),
            total_amount=Decimal("80"),
            gross_amount=Decimal("80"),
            net_amount=Decimal("68.97"),
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            total_quantity=Decimal("7"),
            gross_amount=Decimal("180"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("180"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("155.18"),
        )
        OperatingFinanceBootstrapService().bootstrap()

        stdout = StringIO()
        call_command("audit_operating_finance_sources", period="2026-03", stdout=stdout)
        payload = json.loads(stdout.getvalue())

        self.assertEqual(payload["commercial_classification"]["venta_costeada_total"], "100.00")
        self.assertEqual(payload["commercial_classification"]["venta_no_receta_total"], "0")
        self.assertEqual(payload["commercial_classification"]["venta_receta_sin_match_total"], "80.00")

    def test_unresolved_caja_ch_para_venta_counts_as_non_recipe_resale_not_candidate(self):
        sucursal = Sucursal.objects.create(codigo="MAT", nombre="Matriz")
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=sucursal)
        point_product = PointProduct.objects.create(
            external_id="PCAJCH",
            sku="0240",
            name="CAJA CH PARA VENTA",
            category="Industrias lec",
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=None,
            sale_date=date(2026, 3, 15),
            quantity=Decimal("3"),
            total_amount=Decimal("45"),
            gross_amount=Decimal("45"),
            net_amount=Decimal("38.79"),
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            total_quantity=Decimal("3"),
            gross_amount=Decimal("45"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("45"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("38.79"),
        )
        OperatingFinanceBootstrapService().bootstrap()

        OperatingFinanceSnapshotService().build_snapshot(period_start=date(2026, 3, 1))

        company_row = EmpresaResultadoMensual.objects.get(periodo=date(2026, 3, 1))
        self.assertEqual(company_row.metadata["venta_costeada_total"], "0")
        self.assertEqual(company_row.metadata["venta_no_receta_total"], "45.00")
        self.assertEqual(company_row.metadata["venta_receta_sin_match_total"], "0")

    def test_audit_counts_caja_ch_para_venta_as_non_recipe_reventa(self):
        sucursal = Sucursal.objects.create(codigo="MAT", nombre="Matriz")
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=sucursal)
        point_product = PointProduct.objects.create(
            external_id="PCAJCH2",
            sku="0240",
            name="CAJA CH PARA VENTA",
            category="Industrias lec",
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=None,
            sale_date=date(2026, 3, 18),
            quantity=Decimal("2"),
            total_amount=Decimal("30"),
            gross_amount=Decimal("30"),
            net_amount=Decimal("25.86"),
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            total_quantity=Decimal("2"),
            gross_amount=Decimal("30"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("30"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("25.86"),
        )
        OperatingFinanceBootstrapService().bootstrap()

        stdout = StringIO()
        call_command("audit_operating_finance_sources", period="2026-03", stdout=stdout)
        payload = json.loads(stdout.getvalue())

        self.assertEqual(payload["commercial_classification"]["venta_costeada_total"], "0")
        self.assertEqual(payload["commercial_classification"]["venta_no_receta_total"], "30.00")
        self.assertEqual(payload["commercial_classification"]["venta_receta_sin_match_total"], "0")
        self.assertEqual(payload["commercial_classification"]["non_recipe_bucket_totals"]["REVENTA"], "30.00")
        self.assertEqual(payload["commercial_classification"]["top_candidate_recipe_rows"], [])
        self.assertEqual(payload["commercial_classification"]["top_non_recipe_products"][0]["product_name"], "CAJA CH PARA VENTA")
        self.assertEqual(payload["commercial_classification"]["top_non_recipe_products"][0]["bucket"], "REVENTA")

    def test_forced_resale_products_always_count_as_non_recipe_not_candidate(self):
        sucursal = Sucursal.objects.create(codigo="MAT", nombre="Matriz")
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=sucursal)
        total_amount = Decimal("0")
        for index, case in enumerate(self.FORCED_RESALE_CASES, start=1):
            product = PointProduct.objects.create(
                external_id=f"PFR{index}",
                sku=case["sku"],
                name=case["name"],
                category=case["category"],
            )
            if case["recipe_mode"]:
                Receta.objects.create(
                    nombre=case["name"],
                    codigo_point=case["sku"],
                    hash_contenido=f"hash_forced_resale_{index}",
                    tipo=Receta.TIPO_PRODUCTO_FINAL,
                    modo_costeo=case["recipe_mode"],
                )
            PointDailySale.objects.create(
                branch=point_branch,
                product=product,
                receta=None,
                sale_date=date(2026, 3, 15),
                quantity=Decimal("1"),
                total_amount=case["amount"],
                gross_amount=case["amount"],
                net_amount=case["amount"],
            )
            total_amount += case["amount"]

        PointMonthlySalesOfficial.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            total_quantity=Decimal(str(len(self.FORCED_RESALE_CASES))),
            gross_amount=total_amount,
            discount_amount=Decimal("0"),
            total_amount=total_amount,
            tax_amount=Decimal("0"),
            net_amount=total_amount,
        )
        OperatingFinanceBootstrapService().bootstrap()

        OperatingFinanceSnapshotService().build_snapshot(period_start=date(2026, 3, 1))

        company_row = EmpresaResultadoMensual.objects.get(periodo=date(2026, 3, 1))
        self.assertEqual(company_row.metadata["venta_costeada_total"], "0")
        self.assertEqual(company_row.metadata["venta_no_receta_total"], "436.00")
        self.assertEqual(company_row.metadata["venta_receta_sin_match_total"], "0")

    def test_audit_counts_forced_resale_products_as_non_recipe_reventa(self):
        sucursal = Sucursal.objects.create(codigo="MAT", nombre="Matriz")
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=sucursal)
        total_amount = Decimal("0")
        for index, case in enumerate(self.FORCED_RESALE_CASES, start=1):
            product = PointProduct.objects.create(
                external_id=f"PFRA{index}",
                sku=case["sku"],
                name=case["name"],
                category=case["category"],
            )
            if case["recipe_mode"]:
                Receta.objects.create(
                    nombre=case["name"],
                    codigo_point=case["sku"],
                    hash_contenido=f"hash_forced_resale_audit_{index}",
                    tipo=Receta.TIPO_PRODUCTO_FINAL,
                    modo_costeo=case["recipe_mode"],
                )
            PointDailySale.objects.create(
                branch=point_branch,
                product=product,
                receta=None,
                sale_date=date(2026, 3, 18),
                quantity=Decimal("1"),
                total_amount=case["amount"],
                gross_amount=case["amount"],
                net_amount=case["amount"],
            )
            total_amount += case["amount"]

        PointMonthlySalesOfficial.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            total_quantity=Decimal(str(len(self.FORCED_RESALE_CASES))),
            gross_amount=total_amount,
            discount_amount=Decimal("0"),
            total_amount=total_amount,
            tax_amount=Decimal("0"),
            net_amount=total_amount,
        )
        OperatingFinanceBootstrapService().bootstrap()

        stdout = StringIO()
        call_command("audit_operating_finance_sources", period="2026-03", stdout=stdout)
        payload = json.loads(stdout.getvalue())

        self.assertEqual(payload["commercial_classification"]["venta_costeada_total"], "0")
        self.assertEqual(payload["commercial_classification"]["venta_no_receta_total"], "436.00")
        self.assertEqual(payload["commercial_classification"]["venta_receta_sin_match_total"], "0")
        self.assertEqual(payload["commercial_classification"]["non_recipe_bucket_totals"]["REVENTA"], "436.00")
        self.assertEqual(payload["commercial_classification"]["top_candidate_recipe_rows"], [])
        reported_names = {row["product_name"] for row in payload["commercial_classification"]["top_non_recipe_products"]}
        self.assertEqual(reported_names, {case["name"] for case in self.FORCED_RESALE_CASES})

    def test_product_business_rule_from_db_counts_as_fixed_reventa(self):
        sucursal = Sucursal.objects.create(codigo="MAT", nombre="Matriz")
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=sucursal)
        point_product = PointProduct.objects.create(
            external_id="PDBRULE",
            sku="DBRULE01",
            name="PRODUCTO REVENTA DESDE BD",
            category="Categoria Variada",
        )
        ProductBusinessRule.objects.create(
            product_name="PRODUCTO REVENTA DESDE BD",
            classification=ProductBusinessRule.CLASSIFICATION_REVENTA,
            is_fixed=True,
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=None,
            sale_date=date(2026, 3, 20),
            quantity=Decimal("4"),
            total_amount=Decimal("120"),
            gross_amount=Decimal("120"),
            net_amount=Decimal("103.45"),
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            total_quantity=Decimal("4"),
            gross_amount=Decimal("120"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("120"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("103.45"),
        )
        OperatingFinanceBootstrapService().bootstrap()

        OperatingFinanceSnapshotService().build_snapshot(period_start=date(2026, 3, 1))

        company_row = EmpresaResultadoMensual.objects.get(periodo=date(2026, 3, 1))
        self.assertEqual(company_row.metadata["venta_costeada_total"], "0")
        self.assertEqual(company_row.metadata["venta_no_receta_total"], "120.00")
        self.assertEqual(company_row.metadata["venta_receta_sin_match_total"], "0")

    def test_audit_prefers_product_business_rule_over_fallback_logic(self):
        sucursal = Sucursal.objects.create(codigo="MAT", nombre="Matriz")
        point_branch = PointBranch.objects.create(external_id="1", name="Matriz", erp_branch=sucursal)
        point_product = PointProduct.objects.create(
            external_id="PDBRULE2",
            sku="DBRULE02",
            name="PRODUCTO REVENTA DESDE BD",
            category="Categoria Variada",
        )
        ProductBusinessRule.objects.create(
            product_name="PRODUCTO REVENTA DESDE BD",
            classification=ProductBusinessRule.CLASSIFICATION_REVENTA,
            is_fixed=True,
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=None,
            sale_date=date(2026, 3, 20),
            quantity=Decimal("4"),
            total_amount=Decimal("120"),
            gross_amount=Decimal("120"),
            net_amount=Decimal("103.45"),
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            total_quantity=Decimal("4"),
            gross_amount=Decimal("120"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("120"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("103.45"),
        )
        OperatingFinanceBootstrapService().bootstrap()

        stdout = StringIO()
        call_command("audit_operating_finance_sources", period="2026-03", stdout=stdout)
        payload = json.loads(stdout.getvalue())

        self.assertEqual(payload["commercial_classification"]["venta_costeada_total"], "0")
        self.assertEqual(payload["commercial_classification"]["venta_no_receta_total"], "120.00")
        self.assertEqual(payload["commercial_classification"]["venta_receta_sin_match_total"], "0")
        self.assertEqual(payload["commercial_classification"]["non_recipe_bucket_totals"]["REVENTA"], "120.00")
        self.assertEqual(payload["commercial_classification"]["top_candidate_recipe_rows"], [])

    def test_product_business_rule_normalizes_name_on_save(self):
        rule = ProductBusinessRule.objects.create(
            product_name="  producto reventa desde bd  ",
            classification=ProductBusinessRule.CLASSIFICATION_REVENTA,
            is_fixed=True,
        )

        self.assertEqual(rule.normalized_name, "PRODUCTO REVENTA DESDE BD")

    def test_product_business_rule_rejects_semantic_duplicate_normalized_name(self):
        ProductBusinessRule.objects.create(
            product_name="PRODUCTO REVENTA DESDE BD",
            classification=ProductBusinessRule.CLASSIFICATION_REVENTA,
            is_fixed=True,
        )

        with self.assertRaises(IntegrityError):
            ProductBusinessRule.objects.create(
                product_name="  producto reventa desde bd  ",
                classification=ProductBusinessRule.CLASSIFICATION_REVENTA,
                is_fixed=True,
            )

    def test_historical_snapshot_deletes_stale_rows_outside_current_scope(self):
        stale_recipe = Receta.objects.create(
            nombre="Sabor Guayaba Grande",
            codigo_point="SGUAYABAG",
            hash_contenido="hist_stale_hash",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            modo_costeo=Receta.MODO_COSTEO_FABRICADO,
        )
        RecetaCostoHistoricoMensual.objects.create(
            periodo=date(2026, 1, 1),
            receta=stale_recipe,
            costo_total=Decimal("0"),
            costo_por_unidad_rendimiento=None,
            lineas_costeadas=0,
            lineas_totales=1,
            coverage_pct=Decimal("0"),
            metadata={"stale": True},
        )

        MonthlyHistoricalCostingService().build_period(period_start=date(2026, 1, 1))

        self.assertFalse(
            RecetaCostoHistoricoMensual.objects.filter(periodo=date(2026, 1, 1), receta=stale_recipe).exists()
        )


class OperatingFinanceExpenseTemplateAndImportTests(TestCase):
    def setUp(self):
        self.sucursal = Sucursal.objects.create(codigo="MAT", nombre="Matriz")
        OperatingFinanceBootstrapService().bootstrap()

    def test_export_template_contains_expected_sheets_and_headers(self):
        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "operating_finance_template.xlsx"
            exported = OperatingFinanceTemplateService().export_template(path)

            self.assertTrue(exported.exists())
            workbook = load_workbook(exported)
            self.assertEqual(workbook.sheetnames, ["Gastos", "Catalogos", "Reglas"])
            headers = [cell.value for cell in workbook["Gastos"][1]]
            self.assertEqual(
                headers,
                [
                    "external_key",
                    "periodo",
                    "centro_costo",
                    "categoria_gasto",
                    "monto",
                    "tipo_dato",
                    "fuente",
                    "es_estimado",
                    "comentario",
                    "archivo_soporte",
                ],
            )

    def test_import_workbook_upserts_by_external_key(self):
        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "operating_finance_template.xlsx"
            OperatingFinanceTemplateService().export_template(path)
            workbook = load_workbook(path)
            sheet = workbook["Gastos"]
            sheet.append(
                [
                    "RENTA_MAT_2026_03",
                    "2026-03",
                    f"SUC_{self.sucursal.codigo}",
                    "RENTA_SUC",
                    "5000.00",
                    GastoOperativoMensual.TIPO_DATO_REAL,
                    GastoOperativoMensual.FUENTE_IMPORTADA,
                    "0",
                    "Renta marzo matriz",
                    "renta_marzo.pdf",
                ]
            )
            workbook.save(path)

            summary = OperatingFinanceExpenseImportService().import_workbook(path)

            self.assertEqual(summary.created, 1)
            self.assertEqual(summary.updated, 0)
            gasto = GastoOperativoMensual.objects.get(external_key="RENTA_MAT_2026_03")
            self.assertEqual(gasto.monto, Decimal("5000.00"))

            workbook = load_workbook(path)
            sheet = workbook["Gastos"]
            sheet["E2"] = "6500.00"
            workbook.save(path)

            summary = OperatingFinanceExpenseImportService().import_workbook(path)

            self.assertEqual(summary.created, 0)
            self.assertEqual(summary.updated, 1)
            gasto.refresh_from_db()
            self.assertEqual(gasto.monto, Decimal("6500.00"))


class GeneralBudgetImportServiceTests(TestCase):
    def _build_budget_workbook(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "GENERAL"
        ws["B1"] = "PRESUPUESTO GENERAL POLLYANA'S DOLCE"
        ws["C3"] = "TOTAL ANUAL"
        ws["F3"] = "ENERO"
        ws["I3"] = "FEBRERO"
        ws["A4"] = "CUENTA"
        ws["B4"] = "CONCEPTO"
        ws["C4"] = "PRESUPUESTO"
        ws["D4"] = "RESULTADO"
        ws["E4"] = "VARIACIÓN"
        ws["F4"] = "PRESUPUESTADO"
        ws["G4"] = "REAL"
        ws["H4"] = "VARIACION"
        ws["I4"] = "PRESUPUESTADO"
        ws["J4"] = "REAL"
        ws["K4"] = "VARIACION"
        ws.append(["4001", "Sueldo", Decimal("120000"), Decimal("118000"), Decimal("0.9833"), Decimal("10000"), Decimal("9800"), Decimal("0.98"), Decimal("10000"), Decimal("9700"), Decimal("0.97")])
        ws.append(["4002", "Renta", Decimal("60000"), Decimal("61000"), Decimal("1.0167"), Decimal("5000"), Decimal("5100"), Decimal("1.02"), Decimal("5000"), Decimal("5050"), Decimal("1.01")])
        wb.save(path)

    def test_import_general_budget_workbook_creates_monthly_lines(self):
        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "PRESUPUESTO ADMINISTRACIÓN 2026.xlsx"
            self._build_budget_workbook(path)

            summary = GeneralBudgetImportService().import_workbook(path)

            self.assertEqual(summary.imports_created, 1)
            self.assertEqual(summary.lines_created, 4)
            self.assertTrue(PresupuestoImport.objects.filter(fuente_nombre=path.name).exists())
            self.assertEqual(PresupuestoLineaMensual.objects.count(), 4)
            enero = PresupuestoLineaMensual.objects.get(period=date(2026, 1, 1), concept="Sueldo")
            self.assertEqual(enero.monthly_budget, Decimal("10000"))
            self.assertEqual(enero.monthly_actual, Decimal("9800"))

    def test_import_general_budget_workbook_upserts_existing_lines(self):
        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "PRESUPUESTO ADMINISTRACIÓN 2026.xlsx"
            self._build_budget_workbook(path)
            GeneralBudgetImportService().import_workbook(path)

            wb = load_workbook(path)
            ws = wb["GENERAL"]
            ws["G5"] = Decimal("9900")
            wb.save(path)

            summary = GeneralBudgetImportService().import_workbook(path)

            self.assertEqual(summary.imports_updated, 1)
            self.assertEqual(summary.lines_updated, 4)
            enero = PresupuestoLineaMensual.objects.get(period=date(2026, 1, 1), concept="Sueldo")
            self.assertEqual(enero.monthly_actual, Decimal("9900"))

    def test_import_general_budget_supports_flat_header_layout(self):
        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "GENERAL"
            ws["A2"] = "PRESUPUESO GENERAL 2026"
            ws["A3"] = "CONCEPTOS"
            ws["B3"] = "ENERO PRESUPUESTADO"
            ws["C3"] = "ENERO REAL"
            ws["D3"] = "VARIACIÓN"
            ws["E3"] = "FEBRERO PRESUPUESTADO"
            ws["F3"] = "FEBRERO REAL"
            ws["G3"] = "VARIACIÓN"
            ws.append(["SUELDO", Decimal("1000"), Decimal("900"), Decimal("0.90"), Decimal("1000"), Decimal("950"), Decimal("0.95")])
            wb.save(path)

            summary = GeneralBudgetImportService().import_workbook(path)

            self.assertEqual(summary.lines_created, 2)
            enero = PresupuestoLineaMensual.objects.get(period=date(2026, 1, 1), concept="SUELDO")
            febrero = PresupuestoLineaMensual.objects.get(period=date(2026, 2, 1), concept="SUELDO")
            self.assertEqual(enero.monthly_budget, Decimal("1000"))
            self.assertEqual(febrero.monthly_actual, Decimal("950"))
            self.assertEqual(enero.monthly_variance, Decimal("0.9"))


class BudgetMonitoringSnapshotServiceTests(TestCase):
    def test_build_snapshot_creates_global_and_source_rows(self):
        import_admin = PresupuestoImport.objects.create(
            tipo=PresupuestoImport.TIPO_GENERAL,
            fuente_nombre="PRESUPUESTO ADMINISTRACIÓN 2026.xlsx",
            archivo_ruta="/tmp/admin.xlsx",
            archivo_hash="hash_admin",
            sheet_name="GENERAL",
            titulo="PRESUPUESTO GENERAL POLLYANA'S DOLCE",
        )
        import_nomina = PresupuestoImport.objects.create(
            tipo=PresupuestoImport.TIPO_GENERAL,
            fuente_nombre="PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx",
            archivo_ruta="/tmp/nomina.xlsx",
            archivo_hash="hash_nomina",
            sheet_name="GENERAL",
            titulo="GENERAL",
        )
        PresupuestoLineaMensual.objects.create(
            importacion=import_admin,
            external_key="admin-enero-1",
            period=date(2026, 1, 1),
            concept="Arrendamiento local",
            monthly_budget=Decimal("1000"),
            monthly_actual=Decimal("900"),
        )
        PresupuestoLineaMensual.objects.create(
            importacion=import_nomina,
            external_key="nomina-enero-1",
            period=date(2026, 1, 1),
            concept="SUELDO",
            monthly_budget=Decimal("500"),
            monthly_actual=Decimal("550"),
        )

        summary = BudgetMonitoringSnapshotService().build_snapshot(period_start=date(2026, 1, 1))

        self.assertEqual(summary.rows_created, 3)
        self.assertEqual(summary.rows_updated, 0)
        self.assertEqual(summary.periods, ["2026-01-01"])

        global_row = PresupuestoResumenMensual.objects.get(
            period=date(2026, 1, 1),
            tipo=PresupuestoResumenMensual.TIPO_GLOBAL,
            fuente_nombre="",
        )
        self.assertEqual(global_row.total_budget, Decimal("1000"))
        self.assertEqual(global_row.total_actual, Decimal("900"))
        self.assertEqual(global_row.line_count, 1)
        self.assertEqual(global_row.metadata["global_mode"], "master_source")

        admin_row = PresupuestoResumenMensual.objects.get(
            period=date(2026, 1, 1),
            tipo=PresupuestoResumenMensual.TIPO_FUENTE,
            fuente_nombre="PRESUPUESTO ADMINISTRACIÓN 2026.xlsx",
        )
        self.assertEqual(admin_row.total_budget, Decimal("1000"))
        self.assertEqual(admin_row.total_actual, Decimal("900"))

    def test_build_snapshot_upserts_existing_rows(self):
        import_obj = PresupuestoImport.objects.create(
            tipo=PresupuestoImport.TIPO_GENERAL,
            fuente_nombre="PRESUPUESTO ADMINISTRACIÓN 2026.xlsx",
            archivo_ruta="/tmp/admin.xlsx",
            archivo_hash="hash_admin",
            sheet_name="GENERAL",
            titulo="PRESUPUESTO GENERAL POLLYANA'S DOLCE",
        )
        PresupuestoLineaMensual.objects.create(
            importacion=import_obj,
            external_key="admin-febrero-1",
            period=date(2026, 2, 1),
            concept="Agua potable",
            monthly_budget=Decimal("100"),
            monthly_actual=Decimal("90"),
        )

        BudgetMonitoringSnapshotService().build_snapshot(period_start=date(2026, 2, 1))

        linea = PresupuestoLineaMensual.objects.get(external_key="admin-febrero-1")
        linea.monthly_actual = Decimal("120")
        linea.save(update_fields=["monthly_actual"])

        summary = BudgetMonitoringSnapshotService().build_snapshot(period_start=date(2026, 2, 1))

        self.assertEqual(summary.rows_created, 0)
        self.assertEqual(summary.rows_updated, 2)
        global_row = PresupuestoResumenMensual.objects.get(
            period=date(2026, 2, 1),
            tipo=PresupuestoResumenMensual.TIPO_GLOBAL,
            fuente_nombre="",
        )
        self.assertEqual(global_row.total_actual, Decimal("120"))

    def test_build_snapshot_falls_back_to_sum_when_no_master_source_exists(self):
        import_sales = PresupuestoImport.objects.create(
            tipo=PresupuestoImport.TIPO_GENERAL,
            fuente_nombre="PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx",
            archivo_ruta="/tmp/ventas.xlsx",
            archivo_hash="hash_ventas",
            sheet_name="GENERAL",
            titulo="GASTOS VENTAS 2026",
        )
        import_nomina = PresupuestoImport.objects.create(
            tipo=PresupuestoImport.TIPO_GENERAL,
            fuente_nombre="PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx",
            archivo_ruta="/tmp/nomina.xlsx",
            archivo_hash="hash_nomina",
            sheet_name="GENERAL",
            titulo="NOMINA GENERAL 2026",
        )
        PresupuestoLineaMensual.objects.create(
            importacion=import_sales,
            external_key="ventas-marzo-1",
            period=date(2026, 3, 1),
            concept="Publicidad",
            monthly_budget=Decimal("700"),
            monthly_actual=Decimal("600"),
        )
        PresupuestoLineaMensual.objects.create(
            importacion=import_nomina,
            external_key="nomina-marzo-1",
            period=date(2026, 3, 1),
            concept="SUELDO",
            monthly_budget=Decimal("300"),
            monthly_actual=Decimal("250"),
        )

        BudgetMonitoringSnapshotService().build_snapshot(period_start=date(2026, 3, 1))

        global_row = PresupuestoResumenMensual.objects.get(
            period=date(2026, 3, 1),
            tipo=PresupuestoResumenMensual.TIPO_GLOBAL,
            fuente_nombre="",
        )
        self.assertEqual(global_row.total_budget, Decimal("1000"))
        self.assertEqual(global_row.total_actual, Decimal("850"))
        self.assertEqual(global_row.line_count, 2)
        self.assertEqual(global_row.metadata["global_mode"], "sum_sources")


class BudgetVsActualServiceTests(TestCase):
    def test_write_example_budget_csv_and_import_is_idempotent(self):
        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "presupuesto_2026_ejemplo.csv"
            exported = write_example_budget_csv(path)
            self.assertEqual(exported.name, "presupuesto_2026_ejemplo.csv")

            text = exported.read_text(encoding="utf-8")
            text = text.replace("ventas,0.00", "ventas,1000.00", 1)
            text = text.replace("utilidad_operativa,0.00", "utilidad_operativa,300.00", 1)
            exported.write_text(text, encoding="utf-8")

            summary = BudgetCsvImportService().import_csv(exported)

            self.assertEqual(summary.lines_created, 132)
            self.assertEqual(summary.lines_updated, 0)
            self.assertEqual(summary.missing_required_concepts, [])
            enero_ventas = PresupuestoLineaMensual.objects.get(
                period=date(2026, 1, 1),
                account_code="ventas",
            )
            self.assertEqual(enero_ventas.monthly_budget, Decimal("1000.00"))

            summary = BudgetCsvImportService().import_csv(exported)

            self.assertEqual(summary.lines_created, 0)
            self.assertEqual(summary.lines_updated, 132)

    def test_budget_vs_actual_snapshot_reads_empresa_resultado_and_persists_summary(self):
        import_obj = PresupuestoImport.objects.create(
            tipo=PresupuestoImport.TIPO_GENERAL,
            fuente_nombre="PRESUPUESTO_2026_CSV",
            archivo_ruta="/tmp/presupuesto.csv",
            archivo_hash="hash-budget",
            sheet_name="CSV_2026",
            titulo="PRESUPUESTO 2026 CSV",
            metadata={"year": 2026},
        )
        rows = {
            "ventas": Decimal("900.00"),
            "costo_mp": Decimal("350.00"),
            "costo_reventa": Decimal("50.00"),
            "gasto_fijo": Decimal("200.00"),
            "mano_obra": Decimal("100.00"),
            "utilidad_operativa": Decimal("250.00"),
        }
        for index, (concept, amount) in enumerate(rows.items(), start=1):
            PresupuestoLineaMensual.objects.create(
                importacion=import_obj,
                external_key=f"budget:{concept}:2026-01",
                period=date(2026, 1, 1),
                account_code=concept,
                concept=concept,
                monthly_budget=amount,
                metadata={"concept_key": concept},
                row_index=index,
            )
        EmpresaResultadoMensual.objects.create(
            periodo=date(2026, 1, 1),
            venta_total=Decimal("1000.00"),
            costo_materia_prima_total=Decimal("300.00"),
            costo_reventa_total=Decimal("40.00"),
            mano_obra_prod_total=Decimal("120.00"),
            indirecto_prod_total=Decimal("30.00"),
            gasto_comercial_total=Decimal("180.00"),
            gasto_corporativo_total=Decimal("20.00"),
            utilidad_operativa_total=Decimal("460.00"),
            metadata={"financial_totals_source": "RENTABILIDAD_SUCURSAL"},
        )

        summary = BudgetVsActualSnapshotService().build_snapshot(period_start=date(2026, 1, 1))

        ventas = next(row for row in summary.rows if row["concept"] == "ventas")
        costo_mp = next(row for row in summary.rows if row["concept"] == "costo_mp")
        self.assertEqual(ventas["variance"], Decimal("100.00"))
        self.assertEqual(ventas["variance_pct"], Decimal("11.11"))
        self.assertEqual(ventas["tone"], "success")
        self.assertEqual(costo_mp["variance"], Decimal("-50.00"))
        self.assertEqual(costo_mp["tone"], "success")

        snapshot = PresupuestoResumenMensual.objects.get(
            period=date(2026, 1, 1),
            tipo=PresupuestoResumenMensual.TIPO_FUENTE,
            fuente_nombre=BUDGET_VS_ACTUAL_SOURCE,
        )
        self.assertEqual(snapshot.total_budget, Decimal("250.00"))
        self.assertEqual(snapshot.total_actual, Decimal("460.00"))
        self.assertEqual(snapshot.metadata["real_source_model"], "reportes.EmpresaResultadoMensual")
        self.assertEqual(snapshot.metadata["presupuesto_fuente"], "LEGACY")
        self.assertEqual(snapshot.metadata["empresa_resultado_financial_source"], "RENTABILIDAD_SUCURSAL")

    def test_budget_vs_actual_prefers_master_budget_lines(self):
        ventas_area = AreaPresupuesto.objects.create(nombre="Ventas", codigo="ventas", orden=1)
        produccion_area = AreaPresupuesto.objects.create(nombre="Producción", codigo="produccion", orden=2)
        compras_area = AreaPresupuesto.objects.create(nombre="Compras", codigo="compras", orden=3)
        admin_area = AreaPresupuesto.objects.create(nombre="Administración", codigo="administracion", orden=3)
        nomina_area = AreaPresupuesto.objects.create(nombre="Nómina", codigo="nomina", orden=4)
        logistica_area = AreaPresupuesto.objects.create(nombre="Logística", codigo="logistica", orden=5)
        gastos_area = AreaPresupuesto.objects.create(nombre="Gastos venta", codigo="gastos-venta", orden=6)
        master_rows = [
            (ventas_area, "Venta total", RubroPresupuesto.TIPO_INGRESO, Decimal("1000.00")),
            (produccion_area, "Costo de producción", RubroPresupuesto.TIPO_COSTO, Decimal("300.00")),
            (produccion_area, "Queso Crema", RubroPresupuesto.TIPO_COSTO, Decimal("900.00")),
            (produccion_area, "Producción indirecta", RubroPresupuesto.TIPO_EGRESO, Decimal("999.00")),
            (compras_area, "Costo reventa", RubroPresupuesto.TIPO_COSTO, Decimal("20.00")),
            (admin_area, "Venta postres", RubroPresupuesto.TIPO_EGRESO, Decimal("500.00")),
            (admin_area, "Venta complementos", RubroPresupuesto.TIPO_EGRESO, Decimal("500.00")),
            (admin_area, "Costos insumos/productos", RubroPresupuesto.TIPO_EGRESO, Decimal("300.00")),
            (admin_area, "Costos complementos", RubroPresupuesto.TIPO_EGRESO, Decimal("20.00")),
            (admin_area, "Arrendamiento local", RubroPresupuesto.TIPO_EGRESO, Decimal("100.00")),
            (nomina_area, "Sueldo", RubroPresupuesto.TIPO_EGRESO, Decimal("80.00")),
            (logistica_area, "Logística", RubroPresupuesto.TIPO_EGRESO, Decimal("40.00")),
            (gastos_area, "Publicidad", RubroPresupuesto.TIPO_EGRESO, Decimal("60.00")),
        ]
        for area, concept, rubro_type, amount in master_rows:
            rubro = RubroPresupuesto.objects.create(area=area, concepto=concept, tipo=rubro_type)
            LineaPresupuestoMensual.objects.create(
                rubro=rubro,
                periodo=date(2026, 1, 1),
                version=LineaPresupuestoMensual.VERSION_ORIGINAL,
                monto_presupuesto=amount,
            )
        EmpresaResultadoMensual.objects.create(
            periodo=date(2026, 1, 1),
            venta_total=Decimal("1100.00"),
            costo_materia_prima_total=Decimal("250.00"),
            mano_obra_prod_total=Decimal("90.00"),
            gasto_comercial_total=Decimal("70.00"),
            gasto_corporativo_total=Decimal("40.00"),
            utilidad_operativa_total=Decimal("650.00"),
        )

        summary = BudgetVsActualSnapshotService().build_snapshot(period_start=date(2026, 1, 1))

        self.assertEqual(summary.budget_source, "MAESTRO")
        self.assertEqual(next(row for row in summary.rows if row["concept"] == "ventas")["budget"], Decimal("1000.00"))
        self.assertEqual(next(row for row in summary.rows if row["concept"] == "costo_mp")["budget"], Decimal("300.00"))
        self.assertEqual(next(row for row in summary.rows if row["concept"] == "costo_reventa")["budget"], Decimal("20.00"))
        self.assertEqual(next(row for row in summary.rows if row["concept"] == "gasto_fijo")["budget"], Decimal("100.00"))
        self.assertEqual(next(row for row in summary.rows if row["concept"] == "nomina")["budget"], Decimal("80.00"))
        self.assertEqual(next(row for row in summary.rows if row["concept"] == "logistica")["budget"], Decimal("40.00"))
        self.assertEqual(next(row for row in summary.rows if row["concept"] == "gastos_venta")["budget"], Decimal("60.00"))
        self.assertEqual(
            next(row for row in summary.rows if row["concept"] == "utilidad_operativa")["budget"],
            Decimal("400.00"),
        )
        snapshot = PresupuestoResumenMensual.objects.get(
            period=date(2026, 1, 1),
            tipo=PresupuestoResumenMensual.TIPO_FUENTE,
            fuente_nombre=BUDGET_VS_ACTUAL_SOURCE,
        )
        self.assertEqual(snapshot.metadata["presupuesto_fuente"], "MAESTRO")
        self.assertEqual(snapshot.metadata["budget_source_model"], "reportes.LineaPresupuestoMensual")


class TrustedBudgetDetailImportServiceTests(TestCase):
    def _build_general_layout_sheet(self, ws, *, title: str, concept: str, annual_budget: Decimal, annual_actual: Decimal, jan_budget: Decimal, jan_actual: Decimal):
        ws["B1"] = title
        ws["C3"] = "TOTAL ANUAL"
        ws["F3"] = "ENERO"
        ws["I3"] = "FEBRERO"
        ws["A4"] = "CUENTA"
        ws["B4"] = "CONCEPTO"
        ws["C4"] = "PRESUPUESTO"
        ws["D4"] = "RESULTADO"
        ws["E4"] = "VARIACIÓN"
        ws["F4"] = "PRESUPUESTADO"
        ws["G4"] = "REAL"
        ws["H4"] = "VARIACION"
        ws["I4"] = "PRESUPUESTADO"
        ws["J4"] = "REAL"
        ws["K4"] = "VARIACION"
        ws.append(["4001", concept, annual_budget, annual_actual, Decimal("0.95"), jan_budget, jan_actual, Decimal("0.98"), jan_budget, jan_actual, Decimal("0.98")])

    def _build_flat_nomina_sheet(self, ws, *, title: str, concept: str, jan_budget: Decimal, jan_actual: Decimal, feb_budget: Decimal, feb_actual: Decimal, ventas_layout: bool = False):
        if ventas_layout:
            ws["B2"] = title
            ws["B3"] = "CONCEPTOS"
            ws["C3"] = "ENERO PRESUPUESTADO"
            ws["D3"] = "ENERO REAL"
            ws["E3"] = "VARIACIÓN"
            ws["F3"] = "FEBRERO PRESUPUESTADO"
            ws["G3"] = "FEBRERO REAL"
            ws["H3"] = "VARIACIÓN"
            ws.append([None, concept, jan_budget, jan_actual, Decimal("0.9"), feb_budget, feb_actual, Decimal("0.95")])
        else:
            ws["A2"] = title
            ws["A3"] = "CONCEPTOS"
            ws["B3"] = "ENERO PRESUPUESTADO"
            ws["C3"] = "ENERO REAL"
            ws["D3"] = "VARIACIÓN"
            ws["E3"] = "FEBRERO PRESUPUESTADO"
            ws["F3"] = "FEBRERO REAL"
            ws["G3"] = "VARIACIÓN"
            ws.append([concept, jan_budget, jan_actual, Decimal("0.9"), feb_budget, feb_actual, Decimal("0.95")])

    def test_import_trusted_budget_detail_folder_imports_configured_sheets(self):
        with TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)

            wb = Workbook()
            ws = wb.active
            ws.title = "ADMON"
            self._build_general_layout_sheet(
                ws,
                title="PRESUPUESTO ADMINISTRACIÓN",
                concept="Arrendamiento local",
                annual_budget=Decimal("120000"),
                annual_actual=Decimal("100000"),
                jan_budget=Decimal("10000"),
                jan_actual=Decimal("9000"),
            )
            wb.save(base / "PRESUPUESTO ADMINISTRACIÓN 2026.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = "GUAMUCHIL"
            self._build_general_layout_sheet(
                ws,
                title="PRESUPUESTO GASTOS DE VENTA",
                concept="Arrendamiento local",
                annual_budget=Decimal("60000"),
                annual_actual=Decimal("50000"),
                jan_budget=Decimal("5000"),
                jan_actual=Decimal("4500"),
            )
            wb.save(base / "PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = "VENTAS"
            self._build_flat_nomina_sheet(
                ws,
                title="PRESUPUESTO GENERAL DE VENTAS 2026",
                concept="SUELDO",
                jan_budget=Decimal("1000"),
                jan_actual=Decimal("900"),
                feb_budget=Decimal("1000"),
                feb_actual=Decimal("950"),
                ventas_layout=True,
            )
            wb.save(base / "PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = "PRESUPUESTO PRODUCCIÓN"
            self._build_general_layout_sheet(
                ws,
                title="PRESUPUESTO PRODUCCIÓN",
                concept="Costo de producción",
                annual_budget=Decimal("240000"),
                annual_actual=Decimal("200000"),
                jan_budget=Decimal("20000"),
                jan_actual=Decimal("18000"),
            )
            wb.save(base / "PRESUPUESTO PRODUCCIÓN 2026 AUTORIZADO.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = "LOGÍSTICA"
            self._build_general_layout_sheet(
                ws,
                title="PRESUPUESTO LOGÍSTICA",
                concept="Combustible",
                annual_budget=Decimal("36000"),
                annual_actual=Decimal("30000"),
                jan_budget=Decimal("3000"),
                jan_actual=Decimal("2500"),
            )
            wb.save(base / "PRESUPUESTO LOGISTICA 2026.xlsx")

            summary = TrustedBudgetDetailImportService().import_folder(base)

            self.assertEqual(summary.imports_created, 5)
            self.assertGreaterEqual(summary.lines_created, 6)
            self.assertTrue(
                PresupuestoImport.objects.filter(
                    tipo=PresupuestoImport.TIPO_DETALLE,
                    fuente_nombre="PRESUPUESTO ADMINISTRACIÓN 2026.xlsx",
                    sheet_name="ADMON",
                ).exists()
            )
            enero = PresupuestoLineaMensual.objects.get(
                importacion__tipo=PresupuestoImport.TIPO_DETALLE,
                importacion__fuente_nombre="PRESUPUESTO ADMINISTRACIÓN 2026.xlsx",
                importacion__sheet_name="ADMON",
                period=date(2026, 1, 1),
                concept="Arrendamiento local",
            )
            self.assertEqual(enero.monthly_budget, Decimal("10000"))
            self.assertEqual(enero.metadata["kind"], "admin_recurrente")

    def test_import_workbook_requires_all_configured_sheets(self):
        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "VENTAS"
            self._build_flat_nomina_sheet(
                ws,
                title="PRESUPUESTO GENERAL DE VENTAS 2026",
                concept="SUELDO",
                jan_budget=Decimal("1000"),
                jan_actual=Decimal("900"),
                feb_budget=Decimal("1000"),
                feb_actual=Decimal("950"),
                ventas_layout=True,
            )
            wb.save(path)

            with self.assertRaisesMessage(ValueError, "no contiene todas las hojas esperadas"):
                TrustedBudgetDetailImportService().import_workbook(path)


class BudgetGeneralAuditServiceTests(TestCase):
    def _build_flat_nomina_sheet(self, ws, *, title: str, concept: str, jan_budget: Decimal, jan_actual: Decimal, feb_budget: Decimal, feb_actual: Decimal, ventas_layout: bool = False):
        if ventas_layout:
            ws["B2"] = title
            ws["B3"] = "CONCEPTOS"
            ws["C3"] = "ENERO PRESUPUESTADO"
            ws["D3"] = "ENERO REAL"
            ws["E3"] = "VARIACIÓN"
            ws["F3"] = "FEBRERO PRESUPUESTADO"
            ws["G3"] = "FEBRERO REAL"
            ws["H3"] = "VARIACIÓN"
            ws.append([None, concept, jan_budget, jan_actual, Decimal("0.9"), feb_budget, feb_actual, Decimal("0.95")])
        else:
            ws["A2"] = title
            ws["A3"] = "CONCEPTOS"
            ws["B3"] = "ENERO PRESUPUESTADO"
            ws["C3"] = "ENERO REAL"
            ws["D3"] = "VARIACIÓN"
            ws["E3"] = "FEBRERO PRESUPUESTADO"
            ws["F3"] = "FEBRERO REAL"
            ws["G3"] = "VARIACIÓN"
            ws.append([concept, jan_budget, jan_actual, Decimal("0.9"), feb_budget, feb_actual, Decimal("0.95")])

    def _build_sales_book(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "GENERAL"
        ws["A3"] = "GASTO"
        ws["C3"] = "TOTAL ANUAL"
        ws["F3"] = "ENERO"
        ws["I3"] = "FEBRERO"
        ws["A4"] = "CUENTA"
        ws["B4"] = "DESCRIPCION"
        ws["C4"] = "PRESUPUESTO"
        ws["D4"] = "RESULTADO"
        ws["E4"] = "VARIACIÓN"
        ws["F4"] = "PRESUPUESTADO"
        ws["G4"] = "REAL"
        ws["H4"] = "VARIACION"
        ws["I4"] = "PRESUPUESTADO"
        ws["J4"] = "REAL"
        ws["K4"] = "VARIACION"
        ws.append([None, "Sueldo", 12000, 0, 0, 1000, 0, 0, 1000, 0, 0])
        ws.append([None, "Arrendamiento local", 2400, 0, 0, 200, 0, 0, 200, 0, 0])
        for name in BudgetGeneralAuditService.SALES_BRANCH_SHEETS:
            branch = wb.create_sheet(name)
            branch["A3"] = "GASTO"
            branch["C3"] = "TOTAL ANUAL"
            branch["F3"] = "ENERO"
            branch["I3"] = "FEBRERO"
            branch["A4"] = "CUENTA"
            branch["B4"] = "DESCRIPCION"
            branch["C4"] = "PRESUPUESTO"
            branch["D4"] = "RESULTADO"
            branch["E4"] = "VARIACIÓN"
            branch["F4"] = "PRESUPUESTADO"
            branch["G4"] = "REAL"
            branch["H4"] = "VARIACION"
            branch["I4"] = "PRESUPUESTADO"
            branch["J4"] = "REAL"
            branch["K4"] = "VARIACION"
            branch.append([None, "Sueldo", 1200, 0, 0, Decimal("111.111111"), 0, 0, Decimal("111.111111"), 0, 0])
            branch.append([None, "Arrendamiento local", 240, 0, 0, Decimal("22.222222"), 0, 0, Decimal("22.222222"), 0, 0])
        wb.save(path)

    def _build_nomina_book(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "GENERAL"
        ws["A2"] = "PRESUPUESO GENERAL 2026"
        ws["A3"] = "CONCEPTOS"
        ws["B3"] = "ENERO PRESUPUESTADO"
        ws["C3"] = "ENERO REAL"
        ws["D3"] = "VARIACIÓN"
        ws["E3"] = "FEBRERO PRESUPUESTADO"
        ws["F3"] = "FEBRERO REAL"
        ws["G3"] = "VARIACIÓN"
        ws.append(["SUELDO", 1000, 0, 0, 1100, 0, 0])
        ws.append(["VACACIONES", 100, 0, 0, 120, 0, 0])
        ws.append(["PRIMA VACACIONES", 20, 0, 0, 24, 0, 0])
        for name in BudgetGeneralAuditService.NOMINA_AREA_SHEETS:
            area = wb.create_sheet(name)
            if name == "VENTAS":
                self._build_flat_nomina_sheet(area, title="VENTAS", concept="SUELDO", jan_budget=250, jan_actual=0, feb_budget=275, feb_actual=0, ventas_layout=True)
                area.append([None, "VACACIONES", 25, 0, 0, 30, 0, 0])
                area.append([None, "PRIMA VACACIONES", 5, 0, 0, 6, 0, 0])
            else:
                self._build_flat_nomina_sheet(area, title=name, concept="SUELDO", jan_budget=250, jan_actual=0, feb_budget=275, feb_actual=0)
                area.append(["VACACIONES", 25, 0, 0, 30, 0, 0])
                area.append(["PRIMA VACACIONES", 5, 0, 0, 6, 0, 0])
        wb.save(path)

    def _build_admin_book(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "GENERAL"
        ws["A3"] = "TOTAL ANUAL"
        ws["F3"] = "ENERO"
        ws["I3"] = "FEBRERO"
        ws["A4"] = "CUENTA"
        ws["B4"] = "DESCRIPCION"
        ws["C4"] = "PRESUPUESTO"
        ws["D4"] = "RESULTADO"
        ws["E4"] = "VARIACIÓN"
        ws["F4"] = "PRESUPUESTADO"
        ws["G4"] = "REAL"
        ws["H4"] = "VARIACION"
        ws["I4"] = "PRESUPUESTADO"
        ws["J4"] = "REAL"
        ws["K4"] = "VARIACION"
        ws["B5"] = "Sueldo"
        ws["F5"] = "=[4]GENERAL!B4"
        ws["I5"] = "=[4]GENERAL!E4"
        ws["B6"] = "Imss"
        ws["F6"] = "=[4]GENERAL!B13"
        ws["I6"] = "=[4]GENERAL!E13"
        ws["B7"] = "Infonavit-RCV"
        ws["F7"] = "=[4]GENERAL!B14"
        ws["I7"] = "=[4]GENERAL!E14"
        ws["B8"] = "Aguinaldo"
        ws["F8"] = "=[4]GENERAL!B23"
        ws["I8"] = "=[4]GENERAL!E23"
        ws["B9"] = "Utilidades"
        ws["F9"] = "=[4]GENERAL!B22"
        ws["I9"] = "=[4]GENERAL!E22"
        wb.save(path)

    def test_budget_general_audit_detects_matching_sales_and_nomina_summaries(self):
        with TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            self._build_sales_book(base / "PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx")
            self._build_nomina_book(base / "PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx")

            payload = BudgetGeneralAuditService().audit_folder(base)

            sales = {row["concept"]: row for row in payload["sales_general_vs_branches"]}
            self.assertAlmostEqual(sales["Sueldo"]["general_enero"], sales["Sueldo"]["detail_enero"], places=4)
            self.assertAlmostEqual(sales["Arrendamiento local"]["general_febrero"], sales["Arrendamiento local"]["detail_febrero"], places=4)
            nomina = {row["concept"]: row for row in payload["nomina_general_vs_areas"]}
            self.assertAlmostEqual(nomina["SUELDO"]["general_enero"], nomina["SUELDO"]["detail_enero"], places=4)
            self.assertAlmostEqual(nomina["PRIMA VACACIONAL"]["general_febrero"], nomina["PRIMA VACACIONAL"]["detail_febrero"], places=4)

    def test_budget_general_audit_flags_admin_general_nomina_reference_mismatches(self):
        with TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            self._build_sales_book(base / "PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx")
            self._build_nomina_book(base / "PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx")
            self._build_admin_book(base / "PRESUPUESTO ADMINISTRACIÓN 2026.xlsx")

            payload = BudgetGeneralAuditService().audit_folder(base)

            findings = {row["concept"]: row for row in payload["admin_general_nomina_links"]}
            self.assertEqual(findings["Sueldo"]["status"], "ok")
            self.assertEqual(findings["Imss"]["status"], "reference_mismatch")
            self.assertEqual(findings["Imss"]["target_label"], "")
            self.assertEqual(findings["Infonavit-RCV"]["status"], "reference_mismatch")
            self.assertEqual(findings["Aguinaldo"]["status"], "reference_mismatch")

    def test_budget_general_audit_runs_full_month_concept_reconciliation(self):
        with TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            self._build_sales_book(base / "PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx")
            self._build_nomina_book(base / "PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx")

            payload = BudgetGeneralAuditService().audit_folder(base)

            sales_full = payload["sales_general_vs_branches_full"]
            nomina_full = payload["nomina_general_vs_areas_full"]
            self.assertEqual(sales_full["mismatch_count"], 0)
            self.assertEqual(sales_full["missing_count"], 0)
            self.assertGreaterEqual(sales_full["ok_count"], 2)
            self.assertEqual(nomina_full["mismatch_count"], 0)
            self.assertEqual(nomina_full["missing_count"], 0)
            self.assertGreaterEqual(nomina_full["ok_count"], 3)

    def test_budget_general_audit_detects_external_value_mismatch(self):
        with TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)
            self._build_sales_book(base / "PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx")
            self._build_nomina_book(base / "PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx")
            self._build_admin_book(base / "PRESUPUESTO ADMINISTRACIÓN 2026.xlsx")

            payload = BudgetGeneralAuditService().audit_folder(base)

            external = payload["admin_general_external_links_full"]
            self.assertGreater(external["reviewed_cells"], 0)
            self.assertGreater(external["missing_source_count"], 0)


class BudgetAuditMaterializationServiceTests(TestCase):
    def test_materialize_marks_detail_and_general_statuses(self):
        with TemporaryDirectory() as tmpdir:
            base = Path(tmpdir)

            sales_wb = Workbook()
            sales_ws = sales_wb.active
            sales_ws.title = "GENERAL"
            sales_ws["A3"] = "GASTO"
            sales_ws["C3"] = "TOTAL ANUAL"
            sales_ws["F3"] = "ENERO"
            sales_ws["A4"] = "CUENTA"
            sales_ws["B4"] = "DESCRIPCION"
            sales_ws["C4"] = "PRESUPUESTO"
            sales_ws["D4"] = "RESULTADO"
            sales_ws["E4"] = "VARIACIÓN"
            sales_ws["F4"] = "PRESUPUESTADO"
            sales_ws["G4"] = "REAL"
            sales_ws["H4"] = "VARIACION"
            sales_ws.append([None, "Publicidad", 1200, 0, 0, 100, 0, 0])
            sales_ws.append([None, "TOTAL GASTOS VENTAS", 1200, 0, 0, 100, 0, 0])
            branch_ws = sales_wb.create_sheet("GUAMUCHIL")
            branch_ws["A3"] = "GASTO"
            branch_ws["C3"] = "TOTAL ANUAL"
            branch_ws["F3"] = "ENERO"
            branch_ws["A4"] = "CUENTA"
            branch_ws["B4"] = "DESCRIPCION"
            branch_ws["C4"] = "PRESUPUESTO"
            branch_ws["D4"] = "RESULTADO"
            branch_ws["E4"] = "VARIACIÓN"
            branch_ws["F4"] = "PRESUPUESTADO"
            branch_ws["G4"] = "REAL"
            branch_ws["H4"] = "VARIACION"
            branch_ws.append([None, "Publicidad", 0, 0, 0, 0, 0, 0])
            sales_wb.save(base / "PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx")

            nomina_wb = Workbook()
            nomina_ws = nomina_wb.active
            nomina_ws.title = "GENERAL"
            nomina_ws["A2"] = "PRESUPUESTO NOMINA"
            nomina_ws["A3"] = "CONCEPTOS"
            nomina_ws["B3"] = "ENERO PRESUPUESTADO"
            nomina_ws.append(["IMSS", 100, 0, 0])
            area_ws = nomina_wb.create_sheet("PRODUCCCION")
            area_ws["A2"] = "PRODUCCION"
            area_ws["A3"] = "CONCEPTOS"
            area_ws["B3"] = "ENERO PRESUPUESTADO"
            area_ws.append(["IMSS", 100, 0, 0])
            nomina_wb.save(base / "PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx")

            admin_wb = Workbook()
            admin_ws = admin_wb.active
            admin_ws.title = "GENERAL"
            admin_ws["A3"] = "TOTAL ANUAL"
            admin_ws["F3"] = "ENERO"
            admin_ws["A4"] = "CUENTA"
            admin_ws["B4"] = "DESCRIPCION"
            admin_ws["F4"] = "PRESUPUESTADO"
            admin_ws["B5"] = "Imss"
            admin_ws["F5"] = "=[4]GENERAL!B13"
            admin_wb.save(base / "PRESUPUESTO ADMINISTRACIÓN 2026.xlsx")

            detail_import = PresupuestoImport.objects.create(
                tipo=PresupuestoImport.TIPO_DETALLE,
                fuente_nombre="PRESUPUESTO PRODUCCIÓN 2026 AUTORIZADO.xlsx",
                sheet_name="PRESUPUESTO PRODUCCIÓN",
                archivo_ruta="/tmp/prod.xlsx",
                archivo_hash="hash-prod-audit",
                metadata={"kind": "production_budget"},
            )
            detail_total = PresupuestoLineaMensual.objects.create(
                importacion=detail_import,
                external_key="detail-total",
                period=date(2026, 1, 1),
                concept="PRODUCCIÓN",
                monthly_budget=Decimal("1000"),
                metadata={"kind": "production_budget"},
            )
            detail_payroll = PresupuestoLineaMensual.objects.create(
                importacion=detail_import,
                external_key="detail-imss",
                period=date(2026, 1, 1),
                concept="IMSS",
                monthly_budget=Decimal("100"),
                metadata={"kind": "production_budget"},
            )
            detail_ok = PresupuestoLineaMensual.objects.create(
                importacion=detail_import,
                external_key="detail-costo",
                period=date(2026, 1, 1),
                concept="Costo de producción",
                monthly_budget=Decimal("500"),
                metadata={"kind": "production_budget"},
            )

            general_sales_import = PresupuestoImport.objects.create(
                tipo=PresupuestoImport.TIPO_GENERAL,
                fuente_nombre="PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx",
                sheet_name="GENERAL",
                archivo_ruta="/tmp/sales.xlsx",
                archivo_hash="hash-sales-general",
            )
            general_sales_line = PresupuestoLineaMensual.objects.create(
                importacion=general_sales_import,
                external_key="general-sales-publicidad",
                period=date(2026, 1, 1),
                concept="Publicidad",
                monthly_budget=Decimal("100"),
            )
            general_admin_import = PresupuestoImport.objects.create(
                tipo=PresupuestoImport.TIPO_GENERAL,
                fuente_nombre="PRESUPUESTO ADMINISTRACIÓN 2026.xlsx",
                sheet_name="GENERAL",
                archivo_ruta="/tmp/admin.xlsx",
                archivo_hash="hash-admin-general",
            )
            general_admin_line = PresupuestoLineaMensual.objects.create(
                importacion=general_admin_import,
                external_key="general-admin-imss",
                period=date(2026, 1, 1),
                concept="Imss",
                monthly_budget=Decimal("0"),
            )

            summary = BudgetAuditMaterializationService().materialize(base)

            detail_total.refresh_from_db()
            detail_payroll.refresh_from_db()
            detail_ok.refresh_from_db()
            general_sales_line.refresh_from_db()
            general_admin_line.refresh_from_db()

            self.assertEqual(detail_total.audit_status, PresupuestoLineaMensual.AUDIT_EXCLUDED_TOTAL)
            self.assertEqual(detail_payroll.audit_status, PresupuestoLineaMensual.AUDIT_EXCLUDED_DUPLICATE)
            self.assertEqual(detail_ok.audit_status, PresupuestoLineaMensual.AUDIT_OK)
            self.assertEqual(general_sales_line.audit_status, PresupuestoLineaMensual.AUDIT_DEVIATION)
            self.assertEqual(general_admin_line.audit_status, PresupuestoLineaMensual.AUDIT_BAD_FORMULA)
            self.assertGreaterEqual(summary.ok_lines, 1)


class ProductionExpenseImportServiceTests(TestCase):
    def setUp(self):
        OperatingFinanceBootstrapService().bootstrap()

    def _build_production_budget_workbook(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "PRESUPUESTO PRODUCCIÓN"
        ws["F3"] = "ENERO"
        ws["I3"] = "FEBRERO"
        ws["F4"] = "PRESUPUESTADO"
        ws["G4"] = "REAL"
        ws["I4"] = "PRESUPUESTADO"
        ws["J4"] = "REAL"
        ws["B5"] = "Servicios publicos"
        ws["B6"] = "Agua potable"
        ws["B7"] = "Energia electrica"
        ws["B8"] = "Gas"
        ws["G6"] = Decimal("100")
        ws["G7"] = Decimal("200")
        ws["G8"] = Decimal("300")
        ws["J6"] = Decimal("150")
        ws["J7"] = Decimal("250")
        ws["J8"] = Decimal("350")
        wb.save(path)

    def _build_payroll_workbook(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "PRODUCCCION"
        ws["A3"] = "CONCEPTOS"
        ws["B3"] = "ENERO PRESUPUESTADO"
        ws["C3"] = "ENERO REAL"
        ws["E3"] = "FEBRERO PRESUPUESTADO"
        ws["F3"] = "FEBRERO REAL"
        ws.append(["SUELDO", Decimal("1000"), Decimal("900"), Decimal("0.9"), Decimal("1000"), Decimal("950")])
        ws.append(["DESCUENTO IMSS", Decimal("10"), Decimal("10"), Decimal("1"), Decimal("10"), Decimal("9")])
        ws.append(["IMSS", Decimal("100"), Decimal("80"), Decimal("0.8"), Decimal("100"), Decimal("90")])
        ws.append(["PLAYERA", Decimal("0"), Decimal("15"), Decimal("0"), Decimal("0"), Decimal("0")])
        ws.append(["TOTAL POR MES", Decimal("0"), Decimal("0"), Decimal("0"), Decimal("0"), Decimal("0")])
        wb.save(path)

    def test_import_folder_materializes_only_supported_production_actuals(self):
        with TemporaryDirectory() as tmpdir:
            folder = Path(tmpdir)
            self._build_production_budget_workbook(folder / "PRESUPUESTO PRODUCCIÓN 2026 AUTORIZADO.xlsx")
            self._build_payroll_workbook(folder / "PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx")

            summary = ProductionExpenseImportService().import_folder(folder)

            self.assertEqual(summary.created, 12)
            self.assertEqual(summary.updated, 0)
            self.assertTrue("DESCUENTO IMSS" in summary.skipped_concepts["PRODUCCCION"])
            self.assertFalse(
                GastoOperativoMensual.objects.filter(comentario="Servicios publicos").exists()
            )
            sueldo = GastoOperativoMensual.objects.get(
                comentario="SUELDO",
                periodo=date(2026, 1, 1),
                categoria_gasto__codigo="MANO_OBRA_PROD",
            )
            self.assertEqual(sueldo.monto, Decimal("900"))
            gas = GastoOperativoMensual.objects.get(
                comentario="Gas",
                periodo=date(2026, 2, 1),
                categoria_gasto__codigo="INDIRECTO_PROD",
            )
            self.assertEqual(gas.monto, Decimal("350"))

    def test_import_folder_upserts_existing_rows(self):
        with TemporaryDirectory() as tmpdir:
            folder = Path(tmpdir)
            production_path = folder / "PRESUPUESTO PRODUCCIÓN 2026 AUTORIZADO.xlsx"
            payroll_path = folder / "PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx"
            self._build_production_budget_workbook(production_path)
            self._build_payroll_workbook(payroll_path)

            ProductionExpenseImportService().import_folder(folder)

            wb = load_workbook(payroll_path)
            ws = wb["PRODUCCCION"]
            ws["C4"] = Decimal("920")
            wb.save(payroll_path)

            summary = ProductionExpenseImportService().import_folder(folder)

            self.assertEqual(summary.created, 0)
            self.assertEqual(summary.updated, 12)
            sueldo = GastoOperativoMensual.objects.get(
                comentario="SUELDO",
                periodo=date(2026, 1, 1),
                categoria_gasto__codigo="MANO_OBRA_PROD",
            )
            self.assertEqual(sueldo.monto, Decimal("920"))

    def test_import_folder_flags_outlier_against_month_budget(self):
        with TemporaryDirectory() as tmpdir:
            folder = Path(tmpdir)
            production_path = folder / "PRESUPUESTO PRODUCCIÓN 2026 AUTORIZADO.xlsx"
            payroll_path = folder / "PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx"
            self._build_production_budget_workbook(production_path)
            self._build_payroll_workbook(payroll_path)

            wb = load_workbook(production_path)
            ws = wb["PRESUPUESTO PRODUCCIÓN"]
            ws["I6"] = Decimal("100")
            ws["J6"] = Decimal("1000")
            wb.save(production_path)

            summary = ProductionExpenseImportService().import_folder(folder)

            self.assertTrue(
                any(
                    item["concept"] == "Agua potable" and item["period"] == "2026-02-01"
                    for item in summary.flagged_outliers
                )
            )

    def test_import_production_workbook_uses_detail_rows_and_corrects_water_decimal_shift(self):
        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "PRESUPUESTO PRODUCCIÓN 2026 AUTORIZADO.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "PRESUPUESTO PRODUCCIÓN"
            ws["F3"] = "ENERO"
            ws["I3"] = "FEBRERO"
            ws["F4"] = "PRESUPUESTADO"
            ws["G4"] = "REAL"
            ws["I4"] = "PRESUPUESTADO"
            ws["J4"] = "REAL"
            ws["B5"] = "PRODUCCIÓN"
            ws["B6"] = "Costo de producción"
            ws["B7"] = None
            ws["B8"] = "Sueldo"
            ws["B24"] = "Servicios publicos"
            ws["B25"] = "Agua potable"
            ws["B26"] = "Energia electrica"
            ws["B29"] = "Gas"
            ws["B43"] = "Mantanimiento equipo/maquinaria"
            ws["G5"] = Decimal("999999")
            ws["G6"] = Decimal("888888")
            ws["G7"] = Decimal("999999")
            ws["G8"] = Decimal("900")
            ws["G25"] = Decimal("100")
            ws["G26"] = Decimal("200")
            ws["G29"] = Decimal("300")
            ws["G43"] = Decimal("400")
            ws["I25"] = Decimal("100")
            ws["J25"] = Decimal("100000")
            wb.save(path)

            summary = ProductionExpenseImportService().import_production_workbook(path, through_month=2)

            self.assertEqual(summary.created, 6)
            self.assertFalse(GastoOperativoMensual.objects.filter(comentario="PRODUCCIÓN").exists())
            self.assertFalse(GastoOperativoMensual.objects.filter(comentario="Costo de producción").exists())
            self.assertFalse(GastoOperativoMensual.objects.filter(monto=Decimal("999999")).exists())
            sueldo = GastoOperativoMensual.objects.get(comentario="Sueldo", periodo=date(2026, 1, 1))
            self.assertEqual(sueldo.categoria_gasto.codigo, "MANO_OBRA_PROD")
            agua_febrero = GastoOperativoMensual.objects.get(comentario="Agua potable", periodo=date(2026, 2, 1))
            self.assertEqual(agua_febrero.monto, Decimal("1000.00"))


class BranchAdminExpenseImportServiceTests(TestCase):
    def setUp(self):
        self.sucursal = Sucursal.objects.create(codigo="GUAMUCHIL", nombre="Guamuchil")
        OperatingFinanceBootstrapService().bootstrap()

    def _build_sales_workbook(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "GUAMUCHIL"
        ws["F3"] = "ENERO"
        ws["I3"] = "FEBRERO"
        ws["F4"] = "PRESUPUESTADO"
        ws["G4"] = "REAL"
        ws["I4"] = "PRESUPUESTADO"
        ws["J4"] = "REAL"
        ws["B5"] = "Sueldo"
        ws["B6"] = "Arrendamiento local"
        ws["B7"] = "Agua potable"
        ws["B8"] = "Telefono e Internet"
        ws["B9"] = "TOTAL GASTOS VENTAS"
        ws["G5"] = Decimal("100")
        ws["G6"] = Decimal("200")
        ws["G7"] = Decimal("50")
        ws["G8"] = Decimal("25")
        ws["J5"] = Decimal("110")
        ws["J6"] = Decimal("210")
        ws["J7"] = Decimal("0")
        ws["J8"] = Decimal("30")
        general = wb.create_sheet("GENERAL")
        general["A1"] = "ignore"
        wb.create_sheet("LISTA GASTOS")
        wb.save(path)

    def _build_admin_workbook(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "GENERAL"
        admon = wb.create_sheet("ADMON")
        admon["F3"] = "ENERO"
        admon["I3"] = "FEBRERO"
        admon["F4"] = "PRESUPUESTADO"
        admon["G4"] = "REAL"
        admon["I4"] = "PRESUPUESTADO"
        admon["J4"] = "REAL"
        admon["B5"] = "Sueldo"
        admon["B6"] = "CONTPAQ"
        admon["B7"] = "INGRESOS"
        admon["G5"] = Decimal("300")
        admon["G6"] = Decimal("40")
        admon["J5"] = Decimal("320")
        admon["J6"] = Decimal("45")
        wb.save(path)

    def test_import_folder_materializes_branch_and_admin_actuals(self):
        with TemporaryDirectory() as tmpdir:
            folder = Path(tmpdir)
            self._build_sales_workbook(folder / "PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx")
            self._build_admin_workbook(folder / "PRESUPUESTO ADMINISTRACIÓN 2026.xlsx")

            summary = BranchAdminExpenseImportService().import_folder(folder)

            self.assertEqual(summary.created, 12)
            self.assertEqual(summary.updated, 0)
            self.assertFalse(summary.flagged_outliers)
            nomina = GastoOperativoMensual.objects.get(
                comentario="Sueldo",
                periodo=date(2026, 1, 1),
                centro_costo__codigo="SUC_GUAMUCHIL",
            )
            self.assertEqual(nomina.categoria_gasto.codigo, "NOMINA_SUC")
            renta = GastoOperativoMensual.objects.get(
                comentario="Arrendamiento local",
                periodo=date(2026, 2, 1),
                centro_costo__codigo="SUC_GUAMUCHIL",
            )
            self.assertEqual(renta.categoria_gasto.codigo, "RENTA_SUC")
            sistemas = GastoOperativoMensual.objects.get(
                comentario="CONTPAQ",
                periodo=date(2026, 1, 1),
                centro_costo__codigo="CORP",
            )
            self.assertEqual(sistemas.categoria_gasto.codigo, "SISTEMAS_CORP")
            self.assertFalse(GastoOperativoMensual.objects.filter(comentario="INGRESOS").exists())

    def test_import_folder_ignores_bimonthly_zero_as_outlier(self):
        with TemporaryDirectory() as tmpdir:
            folder = Path(tmpdir)
            self._build_sales_workbook(folder / "PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx")
            self._build_admin_workbook(folder / "PRESUPUESTO ADMINISTRACIÓN 2026.xlsx")

            wb = load_workbook(folder / "PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx")
            ws = wb["GUAMUCHIL"]
            ws["I7"] = Decimal("10")
            ws["J7"] = Decimal("1000")
            wb.save(folder / "PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx")

            summary = BranchAdminExpenseImportService().import_folder(folder)

            self.assertFalse(
                any(item["concept"] == "Agua potable" for item in summary.flagged_outliers)
            )

    def test_import_sales_workbook_supports_historical_target_year(self):
        with TemporaryDirectory() as tmpdir:
            workbook = Path(tmpdir) / "PRESUPUESTO DE GASTOS VENTAS diciembre 2025.xlsx"
            self._build_sales_workbook(workbook)

            summary = BranchAdminExpenseImportService(
                target_year=2025,
                external_prefix="OPEX_HIST",
            ).import_sales_workbook(workbook)

            self.assertEqual(summary.created, 8)
            self.assertEqual(summary.updated, 0)
            self.assertIn("GUAMUCHIL", summary.affected_branches)
            nomina = GastoOperativoMensual.objects.get(
                comentario="Sueldo",
                periodo=date(2025, 1, 1),
                centro_costo__codigo="SUC_GUAMUCHIL",
            )
            self.assertEqual(nomina.tipo_dato, GastoOperativoMensual.TIPO_DATO_REAL)
            self.assertTrue(nomina.external_key.startswith("OPEX_HIST|"))


class HistoricalBranchExpenseImportServiceTests(TestCase):
    def setUp(self):
        self.sucursal = Sucursal.objects.create(codigo="GUAMUCHIL", nombre="Guamuchil")
        OperatingFinanceBootstrapService().bootstrap()

    def _build_sales_workbook(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "GUAMUCHIL"
        ws["F3"] = "ENERO"
        ws["I3"] = "FEBRERO"
        ws["F4"] = "PRESUPUESTADO"
        ws["G4"] = "REAL"
        ws["I4"] = "PRESUPUESTADO"
        ws["J4"] = "REAL"
        ws["B5"] = "Sueldo"
        ws["B6"] = "Arrendamiento local"
        ws["G5"] = Decimal("100")
        ws["G6"] = Decimal("200")
        ws["J5"] = Decimal("110")
        ws["J6"] = Decimal("210")
        general = wb.create_sheet("GENERAL")
        general["A1"] = "ignore"
        wb.create_sheet("LISTA GASTOS")
        wb.save(path)

    def test_import_sales_history_workbook_creates_historical_traceability(self):
        with TemporaryDirectory() as tmpdir:
            workbook = Path(tmpdir) / "PRESUPUESTO DE GASTOS VENTAS diciembre 2025.xlsx"
            self._build_sales_workbook(workbook)

            result = HistoricalBranchExpenseImportService().import_sales_history_workbook(
                workbook,
                target_year=2025,
            )

            upload = result.upload
            self.assertEqual(upload.status, CargaGastoOperativoArchivo.STATUS_SUCCESS)
            self.assertEqual(upload.target_year, 2025)
            self.assertEqual(result.classification, HistoricalBranchExpenseImportService.CLASSIFICATION_HISTORICAL_REAL)
            self.assertEqual(upload.loaded_rows, 4)
            self.assertEqual(upload.project_refresh_count, 0)
            self.assertEqual(upload.affected_branches, ["GUAMUCHIL"])
            self.assertEqual(upload.covered_periods, ["2025-01-01", "2025-02-01"])
            self.assertEqual(upload.summary["classification"], "historical_real")
            self.assertEqual(
                GastoOperativoMensual.objects.filter(periodo__year=2025, centro_costo__codigo="SUC_GUAMUCHIL").count(),
                4,
            )
            self.assertFalse(GastoOperativoMensual.objects.filter(periodo__year=2026).exists())


class BranchRealOperatingExpenseImportServiceTests(TestCase):
    def setUp(self):
        self.sucursal = Sucursal.objects.create(codigo="GUAMUCHIL", nombre="Guamuchil")
        OperatingFinanceBootstrapService().bootstrap()

    def _build_workbook(self, path: Path, rows: list[list[object]]):
        wb = Workbook()
        ws = wb.active
        ws.title = "GastosSucursal"
        ws.append(
            [
                "sucursal",
                "periodo",
                "monto",
                "tipo_dato",
                "categoria_gasto",
                "comentario",
            ]
        )
        for row in rows:
            ws.append(row)
        wb.save(path)

    def test_import_workbook_rejects_cross_year_file_without_partial_persist(self):
        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "gasto_real_sucursales.xlsx"
            self._build_workbook(
                path,
                [
                    ["GUAMUCHIL", "2026-01-01", "15000", "REAL", "", "Enero"],
                    ["GUAMUCHIL", "2025-12-01", "14000", "REAL", "", "Histórico"],
                ],
            )

            with self.assertRaises(BranchRealOperatingExpenseImportValidationError) as ctx:
                BranchRealOperatingExpenseImportService().import_workbook(path, refresh_projects=False)

            self.assertEqual(ctx.exception.summary.processed_rows, 2)
            self.assertFalse(GastoOperativoMensual.objects.exists())

    def test_import_workbook_ignores_budget_rows_by_policy(self):
        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "gasto_real_sucursales.xlsx"
            self._build_workbook(
                path,
                [
                    ["GUAMUCHIL", "2026-01-01", "15000", "REAL", "", "Enero"],
                    ["GUAMUCHIL", "2026-02-01", "16000", "PRESUPUESTO", "", "Febrero presupuesto"],
                ],
            )

            summary = BranchRealOperatingExpenseImportService().import_workbook(path, refresh_projects=False)

            self.assertEqual(summary.processed_rows, 2)
            self.assertEqual(summary.loaded_rows, 1)
            self.assertEqual(summary.skipped_non_real, 1)
            gasto = GastoOperativoMensual.objects.get(periodo=date(2026, 1, 1))
            self.assertEqual(gasto.tipo_dato, GastoOperativoMensual.TIPO_DATO_REAL)
            self.assertEqual(gasto.categoria_gasto.codigo, "OPEX_TOTAL_SUC")

    def test_import_workbook_refreshes_project_snapshots_for_affected_branch(self):
        project = ProyectoInversion.objects.create(
            nombre_proyecto="Guamuchil expansión 2026",
            tipo_proyecto=ProyectoInversion.TIPO_APERTURA_SUCURSAL,
            sucursal_relacionada=self.sucursal,
            fecha_inicio=date(2026, 1, 5),
            fecha_apertura=date(2026, 2, 1),
            monto_inversion_planeado=Decimal("300000"),
        )
        FactVentaDiaria.objects.create(
            fecha=date(2026, 2, 14),
            sucursal=self.sucursal,
            producto_clave="PASTEL-UNO",
            producto_nombre="Pastel Uno",
            cantidad=Decimal("100"),
            tickets=25,
            venta_bruta=Decimal("20000"),
            descuento=Decimal("0"),
            venta_total=Decimal("20000"),
            venta_neta=Decimal("17241.38"),
            costo_estimado=Decimal("9000"),
            margen=Decimal("11000"),
            source_kind=FactVentaDiaria.SOURCE_AUTHORITATIVE,
        )

        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "gasto_real_sucursales.xlsx"
            self._build_workbook(
                path,
                [
                    ["GUAMUCHIL", "2026-01-01", "12000", "REAL", "", "Enero real"],
                    ["GUAMUCHIL", "2026-02-01", "18000", "REAL", "", "Febrero real"],
                ],
            )

            summary = BranchRealOperatingExpenseImportService().import_workbook(
                path,
                refresh_projects=True,
                refresh_until=date(2026, 2, 28),
            )

        self.assertEqual(summary.loaded_rows, 2)
        self.assertEqual(summary.project_refresh_count, 1)
        self.assertEqual(summary.project_ids, [project.id])
        snapshot = ProyectoInversionSnapshotMensual.objects.get(
            proyecto=project,
            periodo=date(2026, 2, 1),
        )
        self.assertEqual(snapshot.gastos_operativos, Decimal("18000.00"))
        self.assertEqual(snapshot.data_source, ProyectoInversionSnapshotMensual.DATA_SOURCE_FACT)
        self.assertEqual(snapshot.confidence_score, 75)
        self.assertEqual(snapshot.fuentes.get("expense_coverage_status"), "PARTIAL")


class OperatingExpenseImportAutomationServiceTests(TestCase):
    def setUp(self):
        self.sucursal = Sucursal.objects.create(codigo="GUAMUCHIL", nombre="Guamuchil")
        OperatingFinanceBootstrapService().bootstrap()
        self.project = ProyectoInversion.objects.create(
            nombre_proyecto="Guamuchil auto import 2026",
            tipo_proyecto=ProyectoInversion.TIPO_APERTURA_SUCURSAL,
            sucursal_relacionada=self.sucursal,
            fecha_inicio=date(2026, 1, 5),
            fecha_apertura=date(2026, 2, 1),
            monto_inversion_planeado=Decimal("300000"),
        )
        FactVentaDiaria.objects.create(
            fecha=date(2026, 2, 14),
            sucursal=self.sucursal,
            producto_clave="PASTEL-UNO",
            producto_nombre="Pastel Uno",
            cantidad=Decimal("100"),
            tickets=25,
            venta_bruta=Decimal("20000"),
            descuento=Decimal("0"),
            venta_total=Decimal("20000"),
            venta_neta=Decimal("17241.38"),
            costo_estimado=Decimal("9000"),
            margen=Decimal("11000"),
            source_kind=FactVentaDiaria.SOURCE_AUTHORITATIVE,
        )
        self.automation_service = OperatingExpenseImportAutomationService()
        self.tempdir = TemporaryDirectory()
        self.addCleanup(self.tempdir.cleanup)
        self.automation_service.storage_root = Path(self.tempdir.name) / "automation-storage"

    def _build_workbook_path(self, path: Path, rows: list[list[object]]) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "GastosSucursal"
        ws.append(
            [
                "sucursal",
                "periodo",
                "monto",
                "tipo_dato",
                "categoria_gasto",
                "comentario",
            ]
        )
        for row in rows:
            ws.append(row)
        wb.save(path)
        return path

    def test_process_uploaded_file_creates_history_and_archives_success(self):
        workbook_path = self._build_workbook_path(
            Path(self.tempdir.name) / "source_web.xlsx",
            [
                ["GUAMUCHIL", "2026-01-01", "12000", "REAL", "", "Enero real"],
                ["GUAMUCHIL", "2026-02-01", "18000", "REAL", "", "Febrero real"],
            ],
        )
        uploaded_file = SimpleUploadedFile(
            "gasto_real_sucursales.xlsx",
            workbook_path.read_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        run = self.automation_service.process_uploaded_file(
            uploaded_file,
            target_year=2026,
            refresh_until=date(2026, 2, 28),
        )

        self.assertEqual(run.status, CargaGastoOperativoArchivo.STATUS_SUCCESS)
        self.assertEqual(run.source_channel, CargaGastoOperativoArchivo.SOURCE_WEB)
        self.assertEqual(run.loaded_rows, 2)
        self.assertEqual(run.created_rows, 2)
        self.assertEqual(run.updated_rows, 0)
        self.assertEqual(run.project_refresh_count, 1)
        self.assertEqual(run.affected_branches, ["GUAMUCHIL"])
        self.assertEqual(run.covered_periods, ["2026-01-01", "2026-02-01"])
        self.assertEqual(run.summary["status"], CargaGastoOperativoArchivo.STATUS_SUCCESS)
        self.assertIn("/processed/", run.stored_file_path)
        self.assertTrue(Path(run.stored_file_path).exists())
        self.assertEqual(GastoOperativoMensual.objects.filter(tipo_dato=GastoOperativoMensual.TIPO_DATO_REAL).count(), 2)
        snapshot = ProyectoInversionSnapshotMensual.objects.get(
            proyecto=self.project,
            periodo=date(2026, 2, 1),
        )
        self.assertEqual(snapshot.gastos_operativos, Decimal("18000.00"))
        self.assertEqual(snapshot.fuentes.get("expense_coverage_status"), "PARTIAL")

    def test_process_directory_marks_duplicate_file_without_reprocessing(self):
        inbox = Path(self.tempdir.name) / "dropbox"
        inbox.mkdir(parents=True, exist_ok=True)
        rows = [["GUAMUCHIL", "2026-03-01", "19000", "REAL", "", "Marzo real"]]
        self._build_workbook_path(inbox / "gastos_marzo.xlsx", rows)

        first_summary = self.automation_service.process_directory(
            inbox,
            refresh_projects=False,
        )

        self.assertEqual(first_summary.processed_files, 1)
        self.assertEqual(first_summary.success_files, 1)
        first_run = CargaGastoOperativoArchivo.objects.get(pk=first_summary.run_ids[0])
        self.assertEqual(first_run.status, CargaGastoOperativoArchivo.STATUS_SUCCESS)
        self.assertTrue((inbox / "processed").exists())
        self.assertEqual(GastoOperativoMensual.objects.filter(periodo=date(2026, 3, 1)).count(), 1)

        self._build_workbook_path(inbox / "gastos_marzo_duplicado.xlsx", rows)

        second_summary = self.automation_service.process_directory(
            inbox,
            refresh_projects=False,
        )

        self.assertEqual(second_summary.processed_files, 1)
        self.assertEqual(second_summary.duplicate_files, 1)
        duplicate_run = CargaGastoOperativoArchivo.objects.get(pk=second_summary.run_ids[0])
        self.assertEqual(duplicate_run.status, CargaGastoOperativoArchivo.STATUS_DUPLICATE)
        self.assertEqual(duplicate_run.loaded_rows, 0)
        self.assertTrue((inbox / "duplicate").exists())
        self.assertEqual(GastoOperativoMensual.objects.filter(periodo=date(2026, 3, 1)).count(), 1)

    def test_duplicate_is_detected_before_importer_execution(self):
        workbook_path = self._build_workbook_path(
            Path(self.tempdir.name) / "source_duplicate.xlsx",
            [
                ["GUAMUCHIL", "2026-01-01", "12000", "REAL", "", "Enero real"],
            ],
        )
        first_upload = SimpleUploadedFile(
            "gasto_real_sucursales.xlsx",
            workbook_path.read_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.automation_service.process_uploaded_file(
            first_upload,
            target_year=2026,
            refresh_until=date(2026, 2, 28),
        )

        second_upload = SimpleUploadedFile(
            "gasto_real_sucursales_copia.xlsx",
            workbook_path.read_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        with patch.object(
            self.automation_service.import_service,
            "import_workbook",
            side_effect=AssertionError("duplicate should not reach importer"),
        ):
            run = self.automation_service.process_uploaded_file(
                second_upload,
                target_year=2026,
                refresh_until=date(2026, 2, 28),
            )

        self.assertEqual(run.status, CargaGastoOperativoArchivo.STATUS_DUPLICATE)
        self.assertEqual(run.loaded_rows, 0)
        self.assertEqual(run.project_refresh_count, 0)
        self.assertIn("/duplicate/", run.stored_file_path)
        self.assertEqual(GastoOperativoMensual.objects.filter(periodo=date(2026, 1, 1)).count(), 1)

    def test_process_uploaded_file_error_does_not_leave_partial_rows(self):
        workbook_path = self._build_workbook_path(
            Path(self.tempdir.name) / "source_invalid.xlsx",
            [
                ["GUAMUCHIL", "2026-01-01", "12000", "REAL", "", "Enero real"],
                ["GUAMUCHIL", "2025-12-01", "9000", "REAL", "", "Histórico inválido"],
            ],
        )
        uploaded_file = SimpleUploadedFile(
            "gasto_real_invalid.xlsx",
            workbook_path.read_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        run = self.automation_service.process_uploaded_file(
            uploaded_file,
            target_year=2026,
            refresh_until=date(2026, 2, 28),
        )

        self.assertEqual(run.status, CargaGastoOperativoArchivo.STATUS_ERROR)
        self.assertEqual(run.loaded_rows, 0)
        self.assertEqual(run.project_refresh_count, 0)
        self.assertTrue(run.error_log)
        self.assertIn("/failed/", run.stored_file_path)
        self.assertFalse(GastoOperativoMensual.objects.filter(centro_costo__sucursal=self.sucursal).exists())

    def test_process_directory_continues_after_invalid_file(self):
        inbox = Path(self.tempdir.name) / "dropbox_resilient"
        inbox.mkdir(parents=True, exist_ok=True)
        (inbox / "gastos_invalidos.txt").write_text("contenido no soportado", encoding="utf-8")
        self._build_workbook_path(
            inbox / "gastos_validos.xlsx",
            [["GUAMUCHIL", "2026-04-01", "21000", "REAL", "", "Abril real"]],
        )

        summary = self.automation_service.process_directory(
            inbox,
            refresh_projects=False,
        )

        self.assertEqual(summary.processed_files, 2)
        self.assertEqual(summary.success_files, 1)
        self.assertEqual(summary.error_files, 1)
        self.assertEqual(len(summary.run_ids), 2)
        self.assertTrue((inbox / "failed").exists())
        self.assertTrue((inbox / "processed").exists())
        self.assertEqual(GastoOperativoMensual.objects.filter(periodo=date(2026, 4, 1)).count(), 1)


class BudgetAreaUploadServiceTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="finanzas", password="pass123")
        self.service = BudgetAreaUploadService()

    def _build_general_upload(self, *, title: str = "PRESUPUESTO GENERAL 2026") -> bytes:
        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "GENERAL"
            ws["B1"] = title
            ws["C3"] = "TOTAL ANUAL"
            ws["F3"] = "ENERO"
            ws["A4"] = "CUENTA"
            ws["B4"] = "CONCEPTO"
            ws["C4"] = "PRESUPUESTO"
            ws["D4"] = "RESULTADO"
            ws["E4"] = "VARIACIÓN"
            ws["F4"] = "PRESUPUESTADO"
            ws["G4"] = "REAL"
            ws["H4"] = "VARIACION"
            ws.append(
                [
                    "4001",
                    "Sueldo",
                    Decimal("120000"),
                    Decimal("118000"),
                    Decimal("0.98"),
                    Decimal("10000"),
                    Decimal("9800"),
                    Decimal("0.98"),
                ]
            )
            wb.save(path)
            return path.read_bytes()

    def test_duplicate_is_detected_before_reimport(self):
        first_upload = SimpleUploadedFile(
            "presupuesto_general.xlsx",
            self._build_general_upload(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        result = self.service.process_uploaded_file(area_key="general", uploaded_file=first_upload, uploaded_by=self.user)
        self.assertEqual(result.status, BudgetAreaUploadService.STATUS_SUCCESS)
        existing_lines = PresupuestoLineaMensual.objects.count()

        second_upload = SimpleUploadedFile(
            "presupuesto_general_duplicado.xlsx",
            self._build_general_upload(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        with patch.object(
            self.service.general_import_service,
            "import_workbook",
            wraps=self.service.general_import_service.import_workbook,
        ) as import_mock:
            duplicate_result = self.service.process_uploaded_file(
                area_key="general",
                uploaded_file=second_upload,
                uploaded_by=self.user,
            )

        self.assertEqual(duplicate_result.status, BudgetAreaUploadService.STATUS_DUPLICATE)
        self.assertEqual(PresupuestoLineaMensual.objects.count(), existing_lines)
        import_mock.assert_not_called()
        self.assertTrue(AuditLog.objects.filter(action="BUDGET_UPLOAD_DUPLICATE").exists())

    def test_failed_upload_does_not_leave_partial_budget_rows(self):
        invalid_upload = SimpleUploadedFile(
            "invalido.xlsx",
            self._build_general_upload(title="SIN ANIO"),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with self.assertRaises(ValueError):
            self.service.process_uploaded_file(area_key="general", uploaded_file=invalid_upload, uploaded_by=self.user)

        self.assertEqual(PresupuestoImport.objects.count(), 0)
        self.assertEqual(PresupuestoLineaMensual.objects.count(), 0)
        failed_log = AuditLog.objects.filter(action="BUDGET_UPLOAD_FAILED").first()
        self.assertIsNotNone(failed_log)
        self.assertEqual(failed_log.payload["area_key"], "general")


class BudgetAreaUploadViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="lectura_presupuesto", password="pass123")
        group, _ = Group.objects.get_or_create(name="LECTURA")
        self.user.groups.add(group)
        self.client.login(username="lectura_presupuesto", password="pass123")

    def _build_admin_budget_file(self) -> bytes:
        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "admin.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "ADMON"
            ws["B1"] = "PRESUPUESTO ADMINISTRACIÓN 2026"
            ws["C3"] = "TOTAL ANUAL"
            ws["F3"] = "ENERO"
            ws["A4"] = "CUENTA"
            ws["B4"] = "CONCEPTO"
            ws["C4"] = "PRESUPUESTO"
            ws["D4"] = "RESULTADO"
            ws["E4"] = "VARIACIÓN"
            ws["F4"] = "PRESUPUESTADO"
            ws["G4"] = "REAL"
            ws["H4"] = "VARIACION"
            ws.append(
                [
                    "4001",
                    "Arrendamiento local",
                    Decimal("120000"),
                    Decimal("100000"),
                    Decimal("0.95"),
                    Decimal("10000"),
                    Decimal("9000"),
                    Decimal("0.90"),
                ]
            )
            wb.save(path)
            return path.read_bytes()

    def test_budget_upload_screen_renders_area_cards(self):
        response = self.client.get(reverse("reportes:presupuesto_importar_por_area"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Carga de presupuesto por área")
        self.assertContains(response, "Dirección / Finanzas")
        self.assertContains(response, "Administración")
        self.assertContains(response, "Ventas por sucursal")
        self.assertContains(response, "Nómina por área")

    def test_budget_upload_screen_processes_area_file_and_updates_history(self):
        upload = SimpleUploadedFile(
            "presupuesto_admin.xlsx",
            self._build_admin_budget_file(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("reportes:presupuesto_importar_por_area"),
            {"area_key": "admin", "budget_file": upload},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Administración")
        self.assertContains(response, "PRESUPUESTO ADMINISTRACIÓN 2026.xlsx")
        import_obj = PresupuestoImport.objects.get(
            tipo=PresupuestoImport.TIPO_DETALLE,
            fuente_nombre="PRESUPUESTO ADMINISTRACIÓN 2026.xlsx",
            sheet_name="ADMON",
        )
        self.assertEqual(import_obj.metadata["upload_area_key"], "admin")
        self.assertEqual(import_obj.metadata["uploaded_by_id"], self.user.id)
        success_log = AuditLog.objects.filter(action="BUDGET_UPLOAD_SUCCESS").first()
        self.assertIsNotNone(success_log)
        self.assertEqual(success_log.payload["area_key"], "admin")
        self.assertContains(response, "PRESUPUESTO ADMINISTRACIÓN 2026.xlsx::ADMON")
