from __future__ import annotations

import csv
from io import BytesIO
from datetime import date
from decimal import Decimal, InvalidOperation
import json

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook

from compras.models import OrdenCompra, RecepcionCompra
from core.models import Sucursal
from core.audit import log_event
from core.access import can_view_reportes
from django.contrib.auth.models import User
from reportes.models import (
    ProyectoInversion,
    ProyectoInversionEscenario,
    ProyectoInversionGasto,
    ProyectoInversionPagoDeuda,
    ProyectoInversionSnapshotMensual,
)
from reportes.services_expansion_decision import ExpansionDecisionService
from reportes.services_expansion_forecast import ExpansionForecastService, recomendar_apertura
from reportes.services_expansion_calibration import ExpansionCalibrationService
from reportes.services_expansion_simulations import ExpansionSimulationRegistryService
from reportes.services_investment_projects import (
    ProyectoInversionScenarioService,
    ProyectoInversionDashboardService,
    ProyectoInversionRefreshService,
)


def _parse_decimal(value: str | None, *, default: str = "0") -> Decimal:
    try:
        return Decimal(str((value or "").strip() or default))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)


def _parse_int(value: str | None, *, default: int = 0) -> int:
    try:
        return int(str(value or "").strip() or default)
    except (TypeError, ValueError):
        return default


def _parse_date(value: str | None) -> date | None:
    raw = (value or "").strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw)
    except ValueError:
        return None


def _decimal_value(value) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0")


def _require_reportes_access(user) -> None:
    if not can_view_reportes(user):
        raise PermissionDenied("No tienes permisos para ver Reportes.")


def _project_module_tabs(active: str) -> list[dict[str, str | bool]]:
    tabs = [
        ("ventas", reverse("reportes:ventas"), "Ventas"),
        ("cierre", reverse("reportes:cierre_producto"), "Cierre producto"),
        ("financiero", reverse("reportes:financiero"), "Financiero"),
        ("consumo", reverse("reportes:consumo"), "Consumo"),
        ("faltantes", reverse("reportes:faltantes"), "Faltantes"),
        ("bi", reverse("reportes:bi"), "BI"),
        ("sucursales", reverse("reportes:proyectos_inversion_sucursales"), "Sucursales"),
        ("comparativo", reverse("reportes:proyectos_inversion_comparativo"), "Comparativo"),
        ("simulador", reverse("reportes:proyectos_inversion_expansion_simulador"), "Simulador expansión"),
        ("calibracion", reverse("reportes:proyectos_inversion_calibracion"), "Calibración"),
    ]
    return [
        {"key": key, "url": url, "label": label, "active": key == active}
        for key, url, label in tabs
    ]


def _expansion_filter_payload(request: HttpRequest) -> dict[str, object]:
    return {
        "fecha_inicio_desde": _parse_date(request.GET.get("fecha_inicio_desde")),
        "fecha_inicio_hasta": _parse_date(request.GET.get("fecha_inicio_hasta")),
        "tipo_proyecto": (request.GET.get("tipo_proyecto") or "").strip(),
        "estatus": (request.GET.get("estatus") or "").strip(),
        "sort_by": (request.GET.get("sort_by") or "health_score").strip(),
    }


def _expansion_sort_options() -> list[tuple[str, str]]:
    return [
        ("health_score", "Health score"),
        ("roi", "ROI"),
        ("payback", "Payback"),
        ("cash_on_cash", "Cash-on-cash"),
        ("recovery_speed", "Velocidad recuperación"),
    ]


def _sort_expansion_rows(rows: list[dict[str, object]], sort_by: str) -> list[dict[str, object]]:
    sort_map = {
        "roi": lambda item: float(item["roi"]),
        "payback": lambda item: -(float(item["payback_real"] or 0)),
        "cash_on_cash": lambda item: float(item["cash_on_cash"]),
        "health_score": lambda item: float(item["health_score"]),
        "recovery_speed": lambda item: float(item["latest_snapshot"].porcentaje_recuperado if item.get("latest_snapshot") else 0),
    }
    return sorted(rows, key=sort_map.get(sort_by, sort_map["health_score"]), reverse=True)


def _decision_display_label(decision: str | None) -> str:
    return {
        "ABRIR": "RECOMENDADO",
        "ESPERAR": "RIESGOSO",
        "NO_ABRIR": "NO ABRIR",
    }.get((decision or "").strip().upper(), decision or "N/D")


def _format_metric_display(value, metric_type: str) -> str:
    if value is None:
        return "N/D"
    decimal_value = _decimal_value(value)
    if metric_type == "currency":
        return f"${decimal_value:,.2f}"
    if metric_type == "percent":
        return f"{decimal_value:,.2f}%"
    if metric_type == "months":
        return f"{decimal_value:,.2f} meses"
    if metric_type == "score":
        return f"{int(decimal_value)}/100"
    return f"{decimal_value:,.2f}"


def _comparison_delta(base_value, projected_value, *, lower_is_better: bool = False) -> dict[str, object]:
    if base_value is None or projected_value is None:
        return {
            "display": "N/D",
            "class_name": "is-neutral",
        }
    base_decimal = _decimal_value(base_value)
    projected_decimal = _decimal_value(projected_value)
    difference = projected_decimal - base_decimal
    if base_decimal == 0:
        delta_pct = None
    else:
        delta_pct = (difference / base_decimal) * Decimal("100")

    if difference == 0:
        class_name = "is-neutral"
    else:
        improved = difference < 0 if lower_is_better else difference > 0
        class_name = "is-positive" if improved else "is-negative"

    if delta_pct is None:
        display = "N/D"
    else:
        sign = "+" if delta_pct > 0 else ""
        display = f"{sign}{delta_pct:,.2f}%"
    return {
        "display": display,
        "class_name": class_name,
    }


def _build_simulator_comparison_rows(forecast_payload: dict[str, object] | None) -> list[dict[str, object]]:
    if not forecast_payload:
        return []
    historical_reference = forecast_payload.get("historical_reference", {}) or {}
    outputs = forecast_payload.get("outputs", {}) or {}
    metric_specs = [
        ("Ventas", historical_reference.get("sales_base"), outputs.get("projected_sales"), "currency", False),
        ("Gasto operativo", historical_reference.get("operating_expenses_base"), outputs.get("projected_operating_expenses"), "currency", True),
        ("Utilidad operativa", historical_reference.get("operating_profit_base"), outputs.get("projected_operating_profit"), "currency", False),
        ("Flujo libre", historical_reference.get("free_cashflow_base"), outputs.get("projected_free_cashflow"), "currency", False),
        ("Payback", historical_reference.get("payback_base"), outputs.get("projected_payback_months"), "months", True),
        ("ROI", historical_reference.get("roi_base"), outputs.get("projected_roi_pct"), "percent", False),
    ]
    rows: list[dict[str, object]] = []
    for label, base_value, projected_value, metric_type, lower_is_better in metric_specs:
        delta = _comparison_delta(base_value, projected_value, lower_is_better=lower_is_better)
        rows.append(
            {
                "label": label,
                "base_display": _format_metric_display(base_value, metric_type),
                "projected_display": _format_metric_display(projected_value, metric_type),
                "delta_display": delta["display"],
                "delta_class": delta["class_name"],
            }
        )
    return rows


def _build_simulator_sensitivity_rows(
    *,
    forecast_service: ExpansionForecastService,
    expansion_context: dict[str, object],
    base_project: ProyectoInversion | None,
    investment_estimate: Decimal | None,
    monthly_rent: Decimal | None,
    sales_adjustment_pct: Decimal | None,
    scenario: str,
) -> list[dict[str, object]]:
    base_adjustment = sales_adjustment_pct if sales_adjustment_pct is not None else Decimal("0")
    scenarios = [
        ("Ventas -10%", base_adjustment - Decimal("10")),
        ("Escenario actual", base_adjustment),
        ("Ventas +10%", base_adjustment + Decimal("10")),
    ]
    rows: list[dict[str, object]] = []
    for label, adjustment in scenarios:
        forecast_payload = forecast_service.forecast(
            base_project=base_project,
            tipo_proyecto=ProyectoInversion.TIPO_APERTURA_SUCURSAL,
            investment_estimate=investment_estimate,
            monthly_rent=monthly_rent,
            sales_adjustment_pct=adjustment,
            scenario=scenario,
            location_reference="",
            user=None,
        )
        recommendation = recomendar_apertura(
            base_project=base_project,
            tipo_proyecto=ProyectoInversion.TIPO_APERTURA_SUCURSAL,
            investment_estimate=investment_estimate,
            monthly_rent=monthly_rent,
            sales_adjustment_pct=adjustment,
            scenario=scenario,
            location_reference="",
            forecast_payload=forecast_payload,
            expansion_context=expansion_context,
            user=None,
        )
        rows.append(
            {
                "label": label,
                "adjustment_pct": adjustment,
                "free_cashflow_display": _format_metric_display(
                    forecast_payload["outputs"].get("projected_free_cashflow"),
                    "currency",
                ),
                "payback_display": _format_metric_display(
                    forecast_payload["outputs"].get("projected_payback_months"),
                    "months",
                ),
                "roi_display": _format_metric_display(
                    forecast_payload["outputs"].get("projected_roi_pct"),
                    "percent",
                ),
                "recommendation_display": _decision_display_label(recommendation.get("decision")),
                "recommendation_class": {
                    "ABRIR": "is-positive",
                    "NO_ABRIR": "is-negative",
                }.get(recommendation.get("decision"), "is-warning"),
            }
        )
    return rows


def _simulator_inputs_from_request(request: HttpRequest) -> dict[str, object]:
    base_project_id = _parse_int(request.POST.get("base_project_id") or request.GET.get("base_project_id")) or None
    scenario = (
        request.POST.get("scenario")
        or request.GET.get("scenario")
        or ExpansionForecastService.SCENARIO_BASE
    ).strip().upper()
    investment_raw = (request.POST.get("investment_estimate") or request.GET.get("investment_estimate") or "").strip()
    monthly_rent_raw = (request.POST.get("monthly_rent_reference") or request.GET.get("monthly_rent_reference") or "").strip()
    sales_adjustment_raw = (request.POST.get("sales_adjustment_reference") or request.GET.get("sales_adjustment_reference") or "").strip()
    return {
        "base_project_id": base_project_id,
        "scenario": scenario,
        "investment_estimate": _parse_decimal(investment_raw) if investment_raw else None,
        "monthly_rent": _parse_decimal(monthly_rent_raw) if monthly_rent_raw else None,
        "sales_adjustment_pct": _parse_decimal(sales_adjustment_raw) if sales_adjustment_raw else None,
        "forecast_form": {
            "base_project_id": str(base_project_id or ""),
            "scenario": scenario,
            "investment_estimate": investment_raw,
            "monthly_rent_reference": monthly_rent_raw,
            "sales_adjustment_reference": sales_adjustment_raw,
        },
    }


def _run_expansion_simulation(
    *,
    forecast_service: ExpansionForecastService,
    expansion_context: dict[str, object],
    base_project: ProyectoInversion | None,
    investment_estimate: Decimal | None,
    monthly_rent: Decimal | None,
    sales_adjustment_pct: Decimal | None,
    scenario: str,
    user=None,
) -> tuple[dict[str, object], dict[str, object], list[dict[str, object]], list[dict[str, object]]]:
    forecast_payload = forecast_service.forecast(
        base_project=base_project,
        tipo_proyecto=ProyectoInversion.TIPO_APERTURA_SUCURSAL,
        investment_estimate=investment_estimate,
        monthly_rent=monthly_rent,
        sales_adjustment_pct=sales_adjustment_pct,
        scenario=scenario,
        location_reference="",
        user=user,
    )
    opening_recommendation = recomendar_apertura(
        base_project=base_project,
        tipo_proyecto=ProyectoInversion.TIPO_APERTURA_SUCURSAL,
        investment_estimate=investment_estimate,
        monthly_rent=monthly_rent,
        sales_adjustment_pct=sales_adjustment_pct,
        scenario=scenario,
        location_reference="",
        forecast_payload=forecast_payload,
        expansion_context=expansion_context,
        user=user,
    )
    comparison_rows = _build_simulator_comparison_rows(forecast_payload)
    sensitivity_rows = _build_simulator_sensitivity_rows(
        forecast_service=forecast_service,
        expansion_context=expansion_context,
        base_project=base_project,
        investment_estimate=investment_estimate,
        monthly_rent=monthly_rent,
        sales_adjustment_pct=sales_adjustment_pct,
        scenario=scenario,
    )
    return forecast_payload, opening_recommendation, comparison_rows, sensitivity_rows


def _simple_pdf_bytes(*, title: str, lines: list[str]) -> bytes:
    def _escape(text: str) -> str:
        return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

    page_width = 792
    page_height = 612
    y = 560
    content_lines = ["BT", "/F1 12 Tf", "36 560 Td"]
    first = True
    for raw in [title, *lines[:40]]:
        text = _escape(raw)
        if first:
            content_lines.append(f"({text}) Tj")
            first = False
        else:
            content_lines.append("T*")
            content_lines.append(f"({text}) Tj")
        y -= 14
        if y < 40:
            break
    content_lines.append("ET")
    content = "\n".join(content_lines).encode("latin-1", errors="replace")

    objects = [
        b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj",
        b"2 0 obj << /Type /Pages /Count 1 /Kids [3 0 R] >> endobj",
        f"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width} {page_height}] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj".encode(),
        b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj",
        b"5 0 obj << /Length " + str(len(content)).encode() + b" >> stream\n" + content + b"\nendstream endobj",
    ]

    output = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(len(output))
        output.extend(obj)
        output.extend(b"\n")
    xref_pos = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode())
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode())
    output.extend(
        f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_pos}\n%%EOF".encode()
    )
    return bytes(output)


def _report_title(report_key: str) -> str:
    return {
        "summary": "Resumen ejecutivo",
        "investment": "Detalle de inversión",
        "performance": "Desempeño mensual",
        "recovery": "Recuperación de inversión",
        "comparison": "Comparativo real vs proyectado",
        "debt": "Estado de deuda",
        "roi": "ROI / Payback / VAN / TIR",
        "calibration": "Reporte de calibración",
    }.get(report_key, "Resumen ejecutivo")


def _simulation_report_payload(
    *,
    forecast_payload: dict[str, object],
    opening_recommendation: dict[str, object],
    comparison_rows: list[dict[str, object]],
    sensitivity_rows: list[dict[str, object]],
    saved_label: str = "",
) -> dict[str, object]:
    historical_reference = forecast_payload.get("historical_reference", {}) or {}
    outputs = forecast_payload.get("outputs", {}) or {}
    inputs = forecast_payload.get("inputs", {}) or {}
    base_project = forecast_payload.get("base_project") or {}
    return {
        "saved_label": saved_label,
        "base_project_name": base_project.get("name") or historical_reference.get("source_project_name") or "Portafolio comparable",
        "scenario_label": forecast_payload.get("scenario_label") or forecast_payload.get("scenario") or "Base",
        "generated_at": forecast_payload.get("generated_at") or timezone.now(),
        "investment_estimate": inputs.get("investment_estimate"),
        "monthly_rent": inputs.get("monthly_rent"),
        "sales_adjustment_pct": inputs.get("sales_adjustment_pct"),
        "projected_sales": outputs.get("projected_sales"),
        "projected_operating_expenses": outputs.get("projected_operating_expenses"),
        "projected_operating_profit": outputs.get("projected_operating_profit"),
        "projected_free_cashflow": outputs.get("projected_free_cashflow"),
        "projected_recovery_cashflow": outputs.get("projected_recovery_cashflow"),
        "projected_payback_months": outputs.get("projected_payback_months"),
        "projected_roi_pct": outputs.get("projected_roi_pct"),
        "projected_health_score": outputs.get("projected_health_score"),
        "recommendation_label": _decision_display_label(opening_recommendation.get("decision")),
        "recommendation_explanation": opening_recommendation.get("explanation") or "",
        "comparison_rows": comparison_rows,
        "sensitivity_rows": sensitivity_rows,
        "decision_rules": opening_recommendation.get("decision_rules", []),
        "assumptions": forecast_payload.get("assumptions", []),
        "historical_reference": historical_reference,
    }


def _simulation_report_lines(report_payload: dict[str, object]) -> list[str]:
    generated_at = report_payload.get("generated_at")
    generated_label = generated_at.strftime("%Y-%m-%d %H:%M") if hasattr(generated_at, "strftime") else str(generated_at or "")
    lines = [
        f"Simulación: {report_payload.get('saved_label') or 'Actual'}",
        f"Fecha: {generated_label}",
        f"Sucursal base: {report_payload.get('base_project_name')}",
        f"Escenario: {report_payload.get('scenario_label')}",
        f"Inversion estimada: {_format_metric_display(report_payload.get('investment_estimate'), 'currency')}",
        f"Renta mensual: {_format_metric_display(report_payload.get('monthly_rent'), 'currency')}",
        f"Ajuste ventas: {_format_metric_display(report_payload.get('sales_adjustment_pct'), 'percent')}",
        "",
        f"Ventas proyectadas: {_format_metric_display(report_payload.get('projected_sales'), 'currency')}",
        f"Gasto operativo proyectado: {_format_metric_display(report_payload.get('projected_operating_expenses'), 'currency')}",
        f"Utilidad operativa: {_format_metric_display(report_payload.get('projected_operating_profit'), 'currency')}",
        f"Flujo libre: {_format_metric_display(report_payload.get('projected_free_cashflow'), 'currency')}",
        f"Flujo para recuperación: {_format_metric_display(report_payload.get('projected_recovery_cashflow'), 'currency')}",
        f"Payback: {_format_metric_display(report_payload.get('projected_payback_months'), 'months')}",
        f"ROI: {_format_metric_display(report_payload.get('projected_roi_pct'), 'percent')}",
        f"Health score: {_format_metric_display(report_payload.get('projected_health_score'), 'score')}",
        "",
        f"Recomendacion final: {report_payload.get('recommendation_label')}",
        f"Motivo: {report_payload.get('recommendation_explanation')}",
    ]
    comparison_rows = report_payload.get("comparison_rows", []) or []
    if comparison_rows:
        lines.append("")
        lines.append("Comparacion contra base")
        for row in comparison_rows:
            lines.append(
                f"{row['label']} | Base {row['base_display']} | Nuevo {row['projected_display']} | Delta {row['delta_display']}"
            )
    sensitivity_rows = report_payload.get("sensitivity_rows", []) or []
    if sensitivity_rows:
        lines.append("")
        lines.append("Sensibilidad comercial")
        for row in sensitivity_rows:
            lines.append(
                f"{row['label']} | Ajuste {_decimal_value(row['adjustment_pct']):,.2f}% | Flujo {row['free_cashflow_display']} | "
                f"Payback {row['payback_display']} | ROI {row['roi_display']} | {row['recommendation_display']}"
            )
    decision_rules = report_payload.get("decision_rules", []) or []
    if decision_rules:
        lines.append("")
        lines.append("Reglas de decision")
        for rule in decision_rules:
            lines.append(
                f"{rule['label']} | {'Cumple' if rule.get('passed') else 'No cumple'} | {rule.get('detail') or ''}"
            )
    assumptions = report_payload.get("assumptions", []) or []
    if assumptions:
        lines.append("")
        lines.append("Supuestos y trazabilidad")
        for item in assumptions[:8]:
            lines.append(f"- {item}")
    return lines


def _export_simulation_pdf(report_payload: dict[str, object]) -> HttpResponse:
    pdf_bytes = _simple_pdf_bytes(
        title=f"Simulacion expansion: {report_payload.get('base_project_name')}",
        lines=_simulation_report_lines(report_payload),
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="simulacion_expansion.pdf"'
    return response


def _export_simulation_xlsx(report_payload: dict[str, object]) -> HttpResponse:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Resumen"
    summary_ws.append(["Simulación", report_payload.get("saved_label") or "Actual"])
    summary_ws.append(["Sucursal base", report_payload.get("base_project_name") or ""])
    summary_ws.append(["Escenario", report_payload.get("scenario_label") or ""])
    summary_ws.append(["Inversión estimada", float(_decimal_value(report_payload.get("investment_estimate")))])
    summary_ws.append(["Renta mensual", float(_decimal_value(report_payload.get("monthly_rent")))])
    summary_ws.append(["Ajuste ventas %", float(_decimal_value(report_payload.get("sales_adjustment_pct")))])
    summary_ws.append(["Ventas proyectadas", float(_decimal_value(report_payload.get("projected_sales")))])
    summary_ws.append(["Gasto operativo proyectado", float(_decimal_value(report_payload.get("projected_operating_expenses")))])
    summary_ws.append(["Utilidad operativa", float(_decimal_value(report_payload.get("projected_operating_profit")))])
    summary_ws.append(["Flujo libre", float(_decimal_value(report_payload.get("projected_free_cashflow")))])
    summary_ws.append(["Flujo para recuperación", float(_decimal_value(report_payload.get("projected_recovery_cashflow")))])
    summary_ws.append(["Payback", float(_decimal_value(report_payload.get("projected_payback_months")))])
    summary_ws.append(["ROI %", float(_decimal_value(report_payload.get("projected_roi_pct")))])
    summary_ws.append(["Health score", int(_decimal_value(report_payload.get("projected_health_score")))])
    summary_ws.append(["Recomendación", report_payload.get("recommendation_label") or ""])
    summary_ws.append(["Explicación", report_payload.get("recommendation_explanation") or ""])

    comparison_ws = wb.create_sheet("Comparacion")
    comparison_ws.append(["Metrica", "Base", "Proyecto nuevo", "Delta"])
    for row in report_payload.get("comparison_rows", []) or []:
        comparison_ws.append([row["label"], row["base_display"], row["projected_display"], row["delta_display"]])

    sensitivity_ws = wb.create_sheet("Sensibilidad")
    sensitivity_ws.append(["Escenario", "Ajuste ventas %", "Flujo libre", "Payback", "ROI", "Recomendacion"])
    for row in report_payload.get("sensitivity_rows", []) or []:
        sensitivity_ws.append(
            [
                row["label"],
                float(_decimal_value(row["adjustment_pct"])),
                row["free_cashflow_display"],
                row["payback_display"],
                row["roi_display"],
                row["recommendation_display"],
            ]
        )

    rules_ws = wb.create_sheet("Reglas")
    rules_ws.append(["Regla", "Cumple", "Detalle"])
    for row in report_payload.get("decision_rules", []) or []:
        rules_ws.append([row["label"], "Si" if row.get("passed") else "No", row.get("detail") or ""])

    trace_ws = wb.create_sheet("Supuestos")
    trace_ws.append(["Tipo", "Valor"])
    for item in report_payload.get("assumptions", []) or []:
        trace_ws.append(["Supuesto", item])
    historical_reference = report_payload.get("historical_reference", {}) or {}
    for key in ["source_project_name", "months_used", "sales_base", "operating_margin_pct", "operating_expenses_base", "debt_service_base"]:
        trace_ws.append([key, historical_reference.get(key)])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="simulacion_expansion.xlsx"'
    return response


def _calibration_report_lines(calibration_context: dict[str, object]) -> list[str]:
    accuracy = calibration_context["accuracy"]
    forecast = calibration_context["forecast_accuracy"]
    lines = [
        f"Precisión total: {float(accuracy['accuracy_pct'] or 0):,.2f}%",
        f"Casos etiquetados: {accuracy['labeled_cases']}",
        f"Casos correctos: {accuracy['accurate_cases']}",
        f"MAE payback: {float(forecast['payback_mae_months'] or 0):,.2f} meses",
    ]
    if calibration_context["data_gaps"]:
        lines.append("")
        lines.append("Brechas de datos")
        for gap in calibration_context["data_gaps"][:8]:
            lines.append(f"- {gap}")
    if calibration_context["incorrect_cases"]:
        lines.append("")
        lines.append("Casos incorrectos")
        for row in calibration_context["incorrect_cases"][:20]:
            lines.append(
                f"{row['project'].nombre_proyecto} | Sistema {row['classification']} | Real {row['real_classification']} | Health {row['health_score']}"
            )
    return lines


def _export_calibration_xlsx(calibration_context: dict[str, object]) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Calibracion"
    accuracy = calibration_context["accuracy"]
    forecast = calibration_context["forecast_accuracy"]
    ws.append(["Metric", "Value"])
    ws.append(["Precision total %", float(accuracy["accuracy_pct"] or 0)])
    ws.append(["Casos etiquetados", accuracy["labeled_cases"]])
    ws.append(["Casos correctos", accuracy["accurate_cases"]])
    ws.append(["MAE payback meses", float(forecast["payback_mae_months"] or 0)])
    ws.append([])
    ws.append(["Proyecto", "Sistema", "Real", "Diferencia", "Health", "ROI", "Payback", "Meses"])
    for row in calibration_context["rows"]:
        ws.append(
            [
                row["project"].nombre_proyecto,
                row["classification"],
                row["real_classification"] or "",
                row["difference"],
                row["health_score"],
                float(row["roi"] or 0),
                float(row["payback_real"] or 0),
                row["recent_month_count"],
            ]
        )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="expansion_calibration.xlsx"'
    output = BytesIO()
    wb.save(output)
    response.write(output.getvalue())
    return response


def _export_calibration_pdf(calibration_context: dict[str, object]) -> HttpResponse:
    pdf_bytes = _simple_pdf_bytes(
        title="Reporte de calibración de expansión",
        lines=_calibration_report_lines(calibration_context),
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="expansion_calibration.pdf"'
    return response


def _project_report_lines(context: dict[str, object], report_key: str) -> list[str]:
    project = context["project"]
    latest = context.get("latest_snapshot")
    kpis = context["kpis"]
    lines = [
        f"Proyecto: {project.nombre_proyecto}",
        f"Estatus: {project.get_estatus_display()}",
        f"Sucursal: {project.sucursal_relacionada.nombre if project.sucursal_relacionada_id else 'Pendiente'}",
        f"Inicio: {project.fecha_inicio.isoformat()}",
        f"Apertura: {project.fecha_apertura.isoformat() if project.fecha_apertura else 'Pendiente'}",
    ]
    if report_key == "summary":
        lines.extend(
            [
                f"Inversión planeada: ${kpis['planned_investment']:,.2f}",
                f"Inversión real: ${kpis['actual_investment']:,.2f}",
                f"Recuperado: ${kpis['cumulative_recovery']:,.2f}",
                f"Saldo pendiente: ${kpis['pending_balance']:,.2f}",
                f"ROI acumulado: {kpis['roi_cumulative']:,.2f}%",
                f"Cash-on-cash: {float(kpis['cash_on_cash'] or 0):,.2f}%",
                f"Health score: {kpis['health_score']}/100",
            ]
        )
    elif report_key == "investment":
        lines.append("")
        lines.append("CAPEX capturado")
        for row in context["expenses"][:20]:
            lines.append(f"{row.fecha.isoformat()} | {row.get_categoria_display()} | ${row.monto_total:,.2f} | {row.descripcion}")
    elif report_key == "performance":
        lines.append("")
        lines.append("Desempeño mensual")
        for row in context["snapshots"][:18]:
            lines.append(
                f"{row.periodo:%Y-%m} | Venta ${row.ventas_mensuales:,.2f} | "
                f"Utilidad operativa ${float(row.utilidad_operativa or 0):,.2f} | "
                f"Flujo libre ${float(row.flujo_libre or 0):,.2f}"
            )
    elif report_key == "recovery":
        lines.append("")
        lines.append("Avance de recuperación")
        for row in context["snapshots"][:18]:
            lines.append(
                f"{row.periodo:%Y-%m} | Recuperación mes ${float(row.monto_recuperacion_mes or 0):,.2f} | "
                f"Acumulado ${float(row.recuperacion_acumulada or 0):,.2f} | "
                f"% {float(row.porcentaje_recuperado or 0):,.2f}"
            )
    elif report_key == "comparison":
        lines.append("")
        lines.append("Escenarios")
        for row in context["scenario_rows"]:
            lines.append(
                f"{row['scenario_name']} | Payback {float(row['payback_months'] or 0):,.2f} meses | "
                f"ROI anual {float(row['annual_roi_pct'] or 0):,.2f}% | "
                f"Cierre {row['projected_close_date'] or 'N/D'}"
            )
    elif report_key == "debt":
        lines.append("")
        lines.append("Servicio de deuda")
        for row in context["debt_payments"][:24]:
            lines.append(
                f"{row.fecha_pago.isoformat()} | Pago ${row.monto_pago:,.2f} | "
                f"Interés ${row.interes_pagado:,.2f} | Capital ${row.capital_amortizado:,.2f} | "
                f"Saldo ${row.saldo_insoluto:,.2f}"
            )
        if latest is not None:
            lines.append(f"Saldo insoluto actual: ${float(latest.saldo_insoluto or 0):,.2f}")
    elif report_key == "roi":
        if latest is not None:
            lines.extend(
                [
                    "",
                    f"ROI mensual: {float(latest.roi_mensual or 0):,.2f}%",
                    f"ROI acumulado: {float(latest.roi_acumulado or 0):,.2f}%",
                    f"ROI anualizado: {float(latest.roi_anualizado or 0):,.2f}%",
                    f"Cash-on-cash: {float(latest.cash_on_cash or 0):,.2f}%",
                    f"Payback real: {float(latest.payback_real_meses or 0):,.2f} meses",
                    f"Payback proyectado: {float(latest.payback_proyectado_meses or 0):,.2f} meses",
                    f"Payback forecast: {float(latest.payback_forecast_meses or 0):,.2f} meses",
                    f"VAN: {float(latest.van or 0):,.2f}",
                    f"TIR: {float(latest.tir or 0):,.2f}%",
                ]
            )
    return lines


def _export_project_xlsx(context: dict[str, object], report_key: str) -> HttpResponse:
    project = context["project"]
    wb = Workbook()
    ws = wb.active
    ws.title = _report_title(report_key)[:31]
    ws.append(["Proyecto", project.nombre_proyecto])
    ws.append(["Estatus", project.get_estatus_display()])
    ws.append(["Sucursal", project.sucursal_relacionada.nombre if project.sucursal_relacionada_id else "Pendiente"])
    ws.append([])

    if report_key == "summary":
        for label, value in context["kpis"].items():
            ws.append([label, float(value) if isinstance(value, Decimal) else value])
    elif report_key == "investment":
        ws.append(["Fecha", "Categoría", "Subcategoría", "Descripción", "Proveedor", "Monto total", "Financiado", "Referencia compra"])
        for row in context["expenses"]:
            ws.append(
                [
                    row.fecha.isoformat(),
                    row.get_categoria_display(),
                    row.subcategoria,
                    row.descripcion,
                    row.proveedor_nombre,
                    float(row.monto_total),
                    "Sí" if row.financiado else "No",
                    row.referencia_compra,
                ]
            )
    elif report_key == "performance":
        ws.append(["Periodo", "Ventas", "Costo venta", "Utilidad bruta", "Gastos operativos", "Utilidad operativa", "Flujo operativo", "Servicio deuda", "Flujo libre"])
        for row in context["snapshots"]:
            ws.append(
                [
                    row.periodo.isoformat(),
                    float(row.ventas_mensuales),
                    float(row.costo_venta_mensual or 0),
                    float(row.utilidad_bruta or 0),
                    float(row.gastos_operativos or 0),
                    float(row.utilidad_operativa or 0),
                    float(row.flujo_operativo or 0),
                    float(row.servicio_deuda or 0),
                    float(row.flujo_libre or 0),
                ]
            )
    elif report_key == "recovery":
        ws.append(["Periodo", "Recuperación mes", "Acumulado", "Saldo pendiente", "% recuperado"])
        for row in context["snapshots"]:
            ws.append(
                [
                    row.periodo.isoformat(),
                    float(row.monto_recuperacion_mes or 0),
                    float(row.recuperacion_acumulada or 0),
                    float(row.saldo_pendiente or 0),
                    float(row.porcentaje_recuperado or 0),
                ]
            )
    elif report_key == "comparison":
        ws.append(["Escenario", "Ventas mensuales", "Flujo neto", "Recuperación mes", "Payback meses", "ROI anual %", "Cierre estimado"])
        for row in context["scenario_rows"]:
            ws.append(
                [
                    row["scenario_name"],
                    float(row["monthly_sales"] or 0),
                    float(row["monthly_net_cashflow"] or 0),
                    float(row["monthly_recovery_amount"] or 0),
                    float(row["payback_months"] or 0),
                    float(row["annual_roi_pct"] or 0),
                    row["projected_close_date"] or "",
                ]
            )
    elif report_key == "debt":
        ws.append(["Fecha pago", "Monto pago", "Interés", "Capital", "Saldo insoluto", "Referencia"])
        for row in context["debt_payments"]:
            ws.append(
                [
                    row.fecha_pago.isoformat(),
                    float(row.monto_pago),
                    float(row.interes_pagado),
                    float(row.capital_amortizado),
                    float(row.saldo_insoluto),
                    row.referencia,
                ]
            )
    elif report_key == "roi":
        ws.append(["Periodo", "ROI mensual %", "ROI acumulado %", "ROI anualizado %", "Cash-on-cash %", "Payback real", "Payback proyectado", "Payback forecast", "VAN", "TIR"])
        for row in context["snapshots"]:
            ws.append(
                [
                    row.periodo.isoformat(),
                    float(row.roi_mensual or 0),
                    float(row.roi_acumulado or 0),
                    float(row.roi_anualizado or 0),
                    float(row.cash_on_cash or 0),
                    float(row.payback_real_meses or 0),
                    float(row.payback_proyectado_meses or 0),
                    float(row.payback_forecast_meses or 0),
                    float(row.van or 0),
                    float(row.tir or 0),
                ]
            )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="proyecto_inversion_{project.pk}_{report_key}.xlsx"'
    return response


def _export_project_pdf(context: dict[str, object], report_key: str) -> HttpResponse:
    project = context["project"]
    pdf_bytes = _simple_pdf_bytes(
        title=f"{_report_title(report_key)} · {project.nombre_proyecto}",
        lines=_project_report_lines(context, report_key),
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="proyecto_inversion_{project.pk}_{report_key}.pdf"'
    return response


@login_required
def proyectos_inversion(request: HttpRequest) -> HttpResponse:
    _require_reportes_access(request.user)

    if request.method == "POST" and request.POST.get("action") == "create_project":
        fecha_inicio = _parse_date(request.POST.get("fecha_inicio"))
        if not fecha_inicio:
            messages.error(request, "La fecha de inicio es obligatoria.")
            return redirect("reportes:proyectos_inversion")
        project = ProyectoInversion.objects.create(
            nombre_proyecto=(request.POST.get("nombre_proyecto") or "").strip() or "Proyecto sin nombre",
            tipo_proyecto=(request.POST.get("tipo_proyecto") or ProyectoInversion.TIPO_APERTURA_SUCURSAL).strip(),
            sucursal_relacionada_id=_parse_int(request.POST.get("sucursal_relacionada_id")) or None,
            fecha_inicio=fecha_inicio,
            fecha_apertura=_parse_date(request.POST.get("fecha_apertura")),
            responsable_id=_parse_int(request.POST.get("responsable_id")) or None,
            estatus=(request.POST.get("estatus") or ProyectoInversion.ESTATUS_PLANEACION).strip(),
            monto_inversion_planeado=_parse_decimal(request.POST.get("monto_inversion_planeado")),
            capital_inicial_aportado=_parse_decimal(request.POST.get("capital_inicial_aportado")),
            deuda_asociada=_parse_decimal(request.POST.get("deuda_asociada")),
            tasa_interes_anual=_parse_decimal(request.POST.get("tasa_interes_anual")),
            plazo_deuda_meses=_parse_int(request.POST.get("plazo_deuda_meses")),
            pago_mensual_deuda_estimado=_parse_decimal(request.POST.get("pago_mensual_deuda_estimado")),
            discount_rate=_parse_decimal(request.POST.get("discount_rate")),
            roi_objetivo=_parse_decimal(request.POST.get("roi_objetivo")),
            payback_objetivo_meses=_parse_int(request.POST.get("payback_objetivo_meses")),
            porcentaje_utilidad_destinado_a_recuperacion=_parse_decimal(
                request.POST.get("porcentaje_utilidad_destinado_a_recuperacion"),
                default="1",
            ),
            recovery_strategy=(request.POST.get("recovery_strategy") or ProyectoInversion.RECOVERY_FULL_NET_CASHFLOW).strip(),
            recovery_percentage=_parse_decimal(request.POST.get("recovery_percentage"), default="1"),
            cierre_por_recuperacion_total=bool(request.POST.get("cierre_por_recuperacion_total", "1")),
            cierre_por_liquidacion_deuda=bool(request.POST.get("cierre_por_liquidacion_deuda")),
            cierre_por_roi_minimo=bool(request.POST.get("cierre_por_roi_minimo")),
            roi_minimo_cierre=_parse_decimal(request.POST.get("roi_minimo_cierre")),
            observaciones=(request.POST.get("observaciones") or "").strip(),
        )
        log_event(
            request.user,
            "CREATE",
            "reportes.ProyectoInversion",
            project.pk,
            payload={"nombre_proyecto": project.nombre_proyecto, "estatus": project.estatus},
        )
        messages.success(request, "Proyecto de inversión creado.")
        return redirect("reportes:proyecto_inversion_detail", project_id=project.pk)

    if request.GET.get("refresh") == "1":
        refresh_service = ProyectoInversionRefreshService()
        for project in ProyectoInversion.objects.exclude(
            estatus__in=[ProyectoInversion.ESTATUS_CERRADO, ProyectoInversion.ESTATUS_CANCELADO]
        ):
            refresh_service.refresh_project(project, user=request.user)
        messages.success(request, "Snapshots de proyectos actualizados.")
        return redirect("reportes:proyectos_inversion")

    context = ProyectoInversionDashboardService().build_portfolio_context()
    expansion_context = ExpansionDecisionService().build_expansion_context(persist=False)
    decisions_by_project_id = {
        row["project"].pk: row for row in expansion_context["decision_rows"]
    }
    partial_data_count = 0
    official_data_count = 0
    for row in context["projects"]:
        latest_snapshot = row.get("latest_snapshot")
        decision_row = decisions_by_project_id.get(row["project"].pk, {})
        row["classification"] = decision_row.get("classification", "VIGILAR")
        row["recommendation"] = decision_row.get("recommendation", "Sin clasificación disponible.")
        row["data_quality_badge"] = "Sin datos"
        row["is_data_partial"] = True
        if latest_snapshot is not None:
            row["data_quality_badge"] = (
                "Datos oficiales"
                if latest_snapshot.data_source == latest_snapshot.DATA_SOURCE_FACT
                and int(latest_snapshot.confidence_score or 0) >= 100
                else "Datos parciales"
            )
            row["is_data_partial"] = row["data_quality_badge"] != "Datos oficiales"
        if row["is_data_partial"]:
            partial_data_count += 1
        else:
            official_data_count += 1
    context.update(
        {
            "module_tabs": _project_module_tabs("sucursales"),
            "project_statuses": ProyectoInversion.ESTATUS_CHOICES,
            "project_types": ProyectoInversion.TIPO_CHOICES,
            "recovery_strategies": ProyectoInversion.RECOVERY_CHOICES,
            "sucursales": Sucursal.objects.order_by("codigo", "nombre"),
            "responsables": User.objects.filter(is_active=True).order_by("first_name", "username"),
            "expansion_summary": expansion_context["global_summary"],
            "expansion_candidate_rows": expansion_context["candidate_rows"][:5],
            "expansion_risk_rows": expansion_context["risk_rows"][:5],
            "expansion_alerts": expansion_context["strategic_alerts"],
            "data_readiness": {
                "official_count": official_data_count,
                "partial_count": partial_data_count,
            },
        }
    )
    return render(request, "reportes/proyectos_inversion_sucursales.html", context)


@login_required
def proyectos_inversion_comparativo(request: HttpRequest) -> HttpResponse:
    _require_reportes_access(request.user)
    fecha_inicio_desde = _parse_date(request.GET.get("fecha_inicio_desde"))
    fecha_inicio_hasta = _parse_date(request.GET.get("fecha_inicio_hasta"))
    tipo_proyecto = (request.GET.get("tipo_proyecto") or "").strip()
    estatus = (request.GET.get("estatus") or "").strip()
    sort_by = (request.GET.get("sort_by") or "roi").strip()

    context = ProyectoInversionDashboardService().build_comparison_context(
        tipo_proyecto=tipo_proyecto,
        estatus=estatus,
        fecha_inicio_desde=fecha_inicio_desde,
        fecha_inicio_hasta=fecha_inicio_hasta,
        sort_by=sort_by,
    )
    context.update(
        {
            "module_tabs": _project_module_tabs("comparativo"),
            "project_statuses": ProyectoInversion.ESTATUS_CHOICES,
            "project_types": ProyectoInversion.TIPO_CHOICES,
            "sort_options": [
                ("roi", "ROI"),
                ("payback", "Payback"),
                ("cash_on_cash", "Cash-on-cash"),
                ("recovery_speed", "Velocidad recuperación"),
                ("health_score", "Health score"),
            ],
        }
    )
    return render(request, "reportes/proyectos_inversion_comparativo.html", context)


@login_required
def proyectos_inversion_calibracion(request: HttpRequest) -> HttpResponse:
    _require_reportes_access(request.user)
    filters = _expansion_filter_payload(request)
    calibration_service = ExpansionCalibrationService()
    calibration_result = None

    export_format = (request.GET.get("export") or "").strip().lower()
    report_key = (request.GET.get("report") or "").strip().lower()
    calibration_context = calibration_service.build_context(
        tipo_proyecto=filters["tipo_proyecto"],
        estatus=filters["estatus"],
        fecha_inicio_desde=filters["fecha_inicio_desde"],
        fecha_inicio_hasta=filters["fecha_inicio_hasta"],
        use_cache=request.method == "GET",
    )
    if report_key == "calibration" and export_format == "xlsx":
        return _export_calibration_xlsx(calibration_context)
    if report_key == "calibration" and export_format == "pdf":
        return _export_calibration_pdf(calibration_context)

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "set_real_classification":
            project = get_object_or_404(ProyectoInversion, pk=_parse_int(request.POST.get("project_id")))
            calibration_service.set_real_classification(
                project,
                (request.POST.get("real_classification") or "").strip(),
                user=request.user,
            )
            messages.success(request, "Clasificación real actualizada.")
            return redirect("reportes:proyectos_inversion_calibracion")
        if action == "save_calibration_settings":
            calibration_service.update_settings(
                payload={
                    "mode_enabled": bool(request.POST.get("mode_enabled")),
                    "health_weights": {
                        "roi": _parse_decimal(request.POST.get("weight_roi"), default="25"),
                        "free_cashflow": _parse_decimal(request.POST.get("weight_free_cashflow"), default="20"),
                        "sales_growth": _parse_decimal(request.POST.get("weight_sales_growth"), default="15"),
                        "capex": _parse_decimal(request.POST.get("weight_capex"), default="20"),
                        "recovery": _parse_decimal(request.POST.get("weight_recovery"), default="20"),
                    },
                    "classification_thresholds": {
                        "expand_min_health_score": _parse_int(request.POST.get("expand_min_health_score"), default=80),
                        "monitor_min_health_score": _parse_int(request.POST.get("monitor_min_health_score"), default=50),
                        "payback_tolerance_ratio": _parse_decimal(request.POST.get("payback_tolerance_ratio"), default="1.00"),
                        "roi_target_factor": _parse_decimal(request.POST.get("roi_target_factor"), default="1.00"),
                    },
                    "forecast": {
                        "preferred_window_months": _parse_int(request.POST.get("preferred_window_months"), default=3),
                        "fallback_window_months": _parse_int(request.POST.get("fallback_window_months"), default=6),
                        "negative_months_mode": (request.POST.get("negative_months_mode") or "fallback_to_fallback_window").strip(),
                    },
                },
                user=request.user,
            )
            messages.success(request, "Parámetros de calibración guardados.")
            return redirect("reportes:proyectos_inversion_calibracion")
        if action == "run_calibration":
            calibration_result = calibration_service.calibrate(user=request.user)
            messages.success(
                request,
                "Calibración ejecutada." if calibration_result.get("applied") else calibration_result.get("reason", "Calibración no aplicada."),
            )
            calibration_context = calibration_service.build_context(
                tipo_proyecto=filters["tipo_proyecto"],
                estatus=filters["estatus"],
                fecha_inicio_desde=filters["fecha_inicio_desde"],
                fecha_inicio_hasta=filters["fecha_inicio_hasta"],
                use_cache=False,
            )

    context = calibration_context.copy()
    context.update(
        {
            "module_tabs": _project_module_tabs("calibracion"),
            "project_statuses": ProyectoInversion.ESTATUS_CHOICES,
            "project_types": ProyectoInversion.TIPO_CHOICES,
            "calibration_context": calibration_context,
            "calibration_result": calibration_result,
            "real_classification_choices": ExpansionCalibrationService.REAL_CLASSIFICATION_CHOICES,
            "filters": {
                "fecha_inicio_desde": filters["fecha_inicio_desde"],
                "fecha_inicio_hasta": filters["fecha_inicio_hasta"],
                "tipo_proyecto": filters["tipo_proyecto"],
                "estatus": filters["estatus"],
            },
        }
    )
    return render(request, "reportes/proyectos_inversion_calibracion.html", context)


@login_required
def proyectos_inversion_expansion_simulador(request: HttpRequest) -> HttpResponse:
    _require_reportes_access(request.user)

    forecast_service = ExpansionForecastService()
    decision_service = ExpansionDecisionService()
    simulation_service = ExpansionSimulationRegistryService()
    expansion_context = decision_service.build_expansion_context(persist=False)
    forecast_payload = None
    opening_recommendation = None
    comparison_rows: list[dict[str, object]] = []
    sensitivity_rows: list[dict[str, object]] = []
    active_saved_simulation = None
    comparison_matrix = None
    selected_compare_ids: list[int] = []
    forecast_form = {
        "base_project_id": "",
        "scenario": ExpansionForecastService.SCENARIO_BASE,
        "investment_estimate": "",
        "monthly_rent_reference": "",
        "sales_adjustment_reference": "",
    }

    load_simulation_id = _parse_int(request.GET.get("load_simulation")) or None
    duplicate_simulation_id = _parse_int(request.GET.get("duplicate_simulation")) or None
    compare_simulations_raw = request.GET.getlist("compare_simulation")
    selected_compare_ids = sorted({_parse_int(value) for value in compare_simulations_raw if _parse_int(value) > 0})

    if load_simulation_id or duplicate_simulation_id:
        target_id = load_simulation_id or duplicate_simulation_id
        saved_scenario = (
            ProyectoInversionEscenario.objects.select_related("proyecto", "proyecto__sucursal_relacionada", "capturado_por")
            .filter(pk=target_id)
            .exclude(simulacion_hash="")
            .first()
        )
        if saved_scenario is not None:
            loaded_context = simulation_service.load_saved_simulation(saved_scenario)
            forecast_payload = loaded_context["forecast_payload"]
            opening_recommendation = loaded_context["opening_recommendation"]
            comparison_rows = loaded_context["comparison_rows"]
            sensitivity_rows = loaded_context["sensitivity_rows"]
            forecast_form = loaded_context["forecast_form"]
            if load_simulation_id:
                active_saved_simulation = simulation_service.serialize_simulation(saved_scenario)
        elif target_id:
            messages.warning(request, "La simulación solicitada ya no está disponible.")

    if selected_compare_ids:
        comparison_matrix = simulation_service.compare_simulations(selected_compare_ids)

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action in {"run_expansion_forecast", "save_current_simulation", "export_current_simulation"}:
            simulation_inputs = _simulator_inputs_from_request(request)
            forecast_form = simulation_inputs["forecast_form"]
            base_project_id = simulation_inputs["base_project_id"]
            base_project = (
                ProyectoInversion.objects.select_related("sucursal_relacionada", "responsable").filter(pk=base_project_id).first()
                if base_project_id
                else None
            )
            forecast_payload, opening_recommendation, comparison_rows, sensitivity_rows = _run_expansion_simulation(
                forecast_service=forecast_service,
                expansion_context=expansion_context,
                base_project=base_project,
                investment_estimate=simulation_inputs["investment_estimate"],
                monthly_rent=simulation_inputs["monthly_rent"],
                sales_adjustment_pct=simulation_inputs["sales_adjustment_pct"],
                scenario=simulation_inputs["scenario"],
                user=request.user,
            )

            if action == "run_expansion_forecast":
                messages.success(request, "Simulación ejecutiva actualizada.")
            elif action == "save_current_simulation":
                if base_project is None:
                    messages.error(request, "Selecciona una sucursal base para guardar la simulación con trazabilidad completa.")
                else:
                    scenario_status = (request.POST.get("scenario_status") or ProyectoInversionEscenario.ESTATUS_EN_REVISION).strip()
                    executive_note = (request.POST.get("executive_note") or "").strip()
                    saved_scenario, created = simulation_service.save_simulation(
                        project=base_project,
                        forecast_payload=forecast_payload,
                        opening_recommendation=opening_recommendation,
                        comparison_rows=comparison_rows,
                        sensitivity_rows=sensitivity_rows,
                        executive_note=executive_note,
                        status=scenario_status,
                        user=request.user,
                    )
                    active_saved_simulation = simulation_service.serialize_simulation(saved_scenario)
                    messages.success(
                        request,
                        "Simulación guardada." if created else "La simulación ya existía; se actualizó sin duplicar registro.",
                    )
            elif action == "export_current_simulation":
                report_payload = _simulation_report_payload(
                    forecast_payload=forecast_payload,
                    opening_recommendation=opening_recommendation,
                    comparison_rows=comparison_rows,
                    sensitivity_rows=sensitivity_rows,
                    saved_label="Actual",
                )
                export_format = (request.POST.get("export_format") or "pdf").strip().lower()
                log_event(
                    request.user,
                    "EXPORT",
                    "reportes.ProyectoInversionEscenario",
                    base_project.pk if base_project else 0,
                    payload={"format": export_format, "mode": "current_simulation"},
                )
                if export_format == "xlsx":
                    return _export_simulation_xlsx(report_payload)
                return _export_simulation_pdf(report_payload)

        if action == "update_saved_simulation_status":
            scenario_id = _parse_int(request.POST.get("saved_simulation_id"))
            saved_scenario = get_object_or_404(
                ProyectoInversionEscenario.objects.select_related("proyecto", "proyecto__sucursal_relacionada", "capturado_por"),
                pk=scenario_id,
            )
            simulation_service.update_status(
                scenario=saved_scenario,
                status=(request.POST.get("scenario_status") or ProyectoInversionEscenario.ESTATUS_EN_REVISION).strip(),
                executive_note=(request.POST.get("executive_note") or "").strip(),
                user=request.user,
            )
            messages.success(request, "Estatus ejecutivo del escenario actualizado.")
            return redirect(f"{reverse('reportes:proyectos_inversion_expansion_simulador')}?load_simulation={saved_scenario.pk}")

    context = expansion_context
    recent_simulations = [simulation_service.serialize_simulation(item) for item in simulation_service.recent_simulations(limit=8)]
    candidate_simulations = [
        item for item in recent_simulations
        if item["status"] in {
            ProyectoInversionEscenario.ESTATUS_CANDIDATO,
            ProyectoInversionEscenario.ESTATUS_APROBADO_PRELIMINAR,
        }
    ][:4]
    latest_import_runs = simulation_service.latest_import_runs(limit=4)
    recent_activity = simulation_service.recent_activity(limit=10)

    export_saved_simulation_id = _parse_int(request.GET.get("export_saved_simulation")) or None
    export_saved_format = (request.GET.get("export_format") or "").strip().lower()
    if export_saved_simulation_id and export_saved_format in {"pdf", "xlsx"}:
        saved_scenario = get_object_or_404(
            ProyectoInversionEscenario.objects.select_related("proyecto", "proyecto__sucursal_relacionada"),
            pk=export_saved_simulation_id,
        )
        saved_payload = simulation_service.serialize_simulation(saved_scenario)
        report_payload = _simulation_report_payload(
            forecast_payload=saved_payload["forecast_payload"],
            opening_recommendation=saved_payload["opening_recommendation"],
            comparison_rows=saved_payload["comparison_rows"],
            sensitivity_rows=saved_payload["sensitivity_rows"],
            saved_label=saved_scenario.nombre,
        )
        log_event(
            request.user,
            "EXPORT",
            "reportes.ProyectoInversionEscenario",
            saved_scenario.pk,
            payload={"format": export_saved_format, "mode": "saved_simulation"},
        )
        if export_saved_format == "xlsx":
            return _export_simulation_xlsx(report_payload)
        return _export_simulation_pdf(report_payload)

    context.update(
        {
            "module_tabs": _project_module_tabs("simulador"),
            "forecast_scenarios": ExpansionForecastService.SCENARIO_CHOICES,
            "base_projects": ProyectoInversion.objects.select_related("sucursal_relacionada").order_by("-fecha_inicio", "-id"),
            "forecast_payload": forecast_payload,
            "opening_recommendation": opening_recommendation,
            "comparison_rows": comparison_rows,
            "sensitivity_rows": sensitivity_rows,
            "active_saved_simulation": active_saved_simulation,
            "recent_simulations": recent_simulations,
            "candidate_simulations": candidate_simulations,
            "comparison_matrix": comparison_matrix,
            "selected_compare_ids": selected_compare_ids,
            "latest_import_runs": latest_import_runs,
            "recent_activity": recent_activity,
            "forecast_form": forecast_form,
            "recommended_bases": context["candidate_rows"][:3],
            "risk_bases": context["risk_rows"][:3],
            "saved_status_choices": ProyectoInversionEscenario.ESTATUS_CHOICES,
        }
    )
    return render(request, "reportes/proyectos_inversion_simulador.html", context)


@login_required
def proyectos_inversion_expansion(request: HttpRequest) -> HttpResponse:
    action = (request.POST.get("action") or "").strip()
    if action in {"set_real_classification", "save_calibration_settings", "run_calibration"}:
        return proyectos_inversion_calibracion(request)
    return proyectos_inversion_expansion_simulador(request)


@login_required
def proyecto_inversion_detail(request: HttpRequest, project_id: int) -> HttpResponse:
    _require_reportes_access(request.user)
    project = get_object_or_404(
        ProyectoInversion.objects.select_related("sucursal_relacionada", "responsable"),
        pk=project_id,
    )
    refresh_service = ProyectoInversionRefreshService()

    needs_initial_refresh = not project.snapshots_mensuales.exists()
    if needs_initial_refresh and project.fecha_apertura and project.fecha_apertura <= timezone.localdate():
        refresh_service.refresh_project(project, user=request.user)

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "add_expense":
            expense = ProyectoInversionGasto.objects.create(
                proyecto=project,
                fecha=_parse_date(request.POST.get("fecha")) or timezone.localdate(),
                categoria=(request.POST.get("categoria") or ProyectoInversionGasto.CATEGORIA_OTROS).strip(),
                subcategoria=(request.POST.get("subcategoria") or "").strip(),
                descripcion=(request.POST.get("descripcion") or "").strip() or "Gasto sin descripción",
                proveedor_id=_parse_int(request.POST.get("proveedor_id")) or None,
                proveedor_nombre=(request.POST.get("proveedor_nombre") or "").strip(),
                monto=_parse_decimal(request.POST.get("monto")),
                iva=_parse_decimal(request.POST.get("iva")),
                monto_total=_parse_decimal(request.POST.get("monto_total")),
                metodo_pago=(request.POST.get("metodo_pago") or "").strip(),
                financiado=bool(request.POST.get("financiado")),
                referencia_compra=(request.POST.get("referencia_compra") or "").strip(),
                referencia_contable=(request.POST.get("referencia_contable") or "").strip(),
                orden_compra_id=_parse_int(request.POST.get("orden_compra_id")) or None,
                recepcion_compra_id=_parse_int(request.POST.get("recepcion_compra_id")) or None,
                evidencia_url=(request.POST.get("evidencia_url") or "").strip(),
                notas=(request.POST.get("notas") or "").strip(),
                capturado_por=request.user,
            )
            log_event(
                request.user,
                "CREATE",
                "reportes.ProyectoInversionGasto",
                expense.pk,
                payload={"project_id": project.pk, "categoria": expense.categoria, "monto_total": str(expense.monto_total)},
            )
            refresh_service.refresh_project(project, user=request.user)
            messages.success(request, "Gasto CAPEX registrado.")
            return redirect("reportes:proyecto_inversion_detail", project_id=project.pk)

        if action == "add_debt_payment":
            payment = ProyectoInversionPagoDeuda.objects.create(
                proyecto=project,
                fecha_pago=_parse_date(request.POST.get("fecha_pago")) or timezone.localdate(),
                monto_pago=_parse_decimal(request.POST.get("monto_pago")),
                interes_pagado=_parse_decimal(request.POST.get("interes_pagado")),
                capital_amortizado=_parse_decimal(request.POST.get("capital_amortizado")),
                saldo_insoluto=_parse_decimal(request.POST.get("saldo_insoluto")),
                referencia=(request.POST.get("referencia") or "").strip(),
                notas=(request.POST.get("notas") or "").strip(),
                capturado_por=request.user,
            )
            log_event(
                request.user,
                "CREATE",
                "reportes.ProyectoInversionPagoDeuda",
                payment.pk,
                payload={"project_id": project.pk, "monto_pago": str(payment.monto_pago)},
            )
            refresh_service.refresh_project(project, user=request.user)
            messages.success(request, "Pago de deuda registrado.")
            return redirect("reportes:proyecto_inversion_detail", project_id=project.pk)

        if action == "save_scenario":
            scenario, created = ProyectoInversionEscenario.objects.update_or_create(
                proyecto=project,
                nombre=(request.POST.get("nombre") or "").strip() or "Escenario",
                defaults={
                    "tipo_escenario": (request.POST.get("tipo_escenario") or ProyectoInversionEscenario.TIPO_BASE).strip(),
                    "ventas_promedio_mensuales": _parse_decimal(request.POST.get("ventas_promedio_mensuales")),
                    "crecimiento_mensual_pct": _parse_decimal(request.POST.get("crecimiento_mensual_pct")),
                    "margen_bruto_pct": _parse_decimal(request.POST.get("margen_bruto_pct")),
                    "gastos_operativos_mensuales": _parse_decimal(request.POST.get("gastos_operativos_mensuales")),
                    "recovery_strategy_override": (request.POST.get("recovery_strategy_override") or "").strip(),
                    "recovery_percentage_override": _parse_decimal(request.POST.get("recovery_percentage_override"))
                    if (request.POST.get("recovery_percentage_override") or "").strip()
                    else None,
                    "horizonte_meses": _parse_int(request.POST.get("horizonte_meses"), default=24),
                    "notas": (request.POST.get("notas") or "").strip(),
                },
            )
            action_name = "CREATE" if created else "UPDATE"
            log_event(
                request.user,
                action_name,
                "reportes.ProyectoInversionEscenario",
                scenario.pk,
                payload={"project_id": project.pk, "tipo_escenario": scenario.tipo_escenario},
            )
            messages.success(request, "Escenario guardado y recalculado.")
            return redirect("reportes:proyecto_inversion_detail", project_id=project.pk)

        if action == "refresh_project":
            refresh_service.refresh_project(project, user=request.user)
            messages.success(request, "Métricas del proyecto actualizadas.")
            return redirect("reportes:proyecto_inversion_detail", project_id=project.pk)

        if action == "close_project" and project.cierre_manual_habilitado:
            project.estatus = ProyectoInversion.ESTATUS_CERRADO
            project.fecha_cierre = _parse_date(request.POST.get("fecha_cierre")) or timezone.localdate()
            project.kpis_cierre = {
                "manual_reason": (request.POST.get("motivo_cierre") or "").strip(),
                "closed_by": request.user.username,
            }
            project.save(update_fields=["estatus", "fecha_cierre", "kpis_cierre", "actualizado_en"])
            log_event(
                request.user,
                "MANUAL_CLOSE",
                "reportes.ProyectoInversion",
                project.pk,
                payload=project.kpis_cierre,
            )
            messages.success(request, "Proyecto cerrado manualmente.")
            return redirect("reportes:proyecto_inversion_detail", project_id=project.pk)

    export_format = (request.GET.get("export") or "").strip().lower()
    report_key = (request.GET.get("report") or "summary").strip().lower()
    context = ProyectoInversionDashboardService().build_detail_context(project)
    context.update(
        {
            "module_tabs": _project_module_tabs("sucursales"),
            "expense_categories": ProyectoInversionGasto.CATEGORIA_CHOICES,
            "scenario_types": ProyectoInversionEscenario.TIPO_CHOICES,
            "recovery_strategies": ProyectoInversion.RECOVERY_CHOICES,
            "ordenes_compra": OrdenCompra.objects.select_related("proveedor").order_by("-fecha_emision", "-id")[:50],
            "recepciones_compra": RecepcionCompra.objects.select_related("orden").order_by("-fecha_recepcion", "-id")[:50],
            "today": timezone.localdate(),
            "report_definitions": [
                {"key": "summary", "label": "Resumen ejecutivo"},
                {"key": "investment", "label": "Detalle de inversión"},
                {"key": "performance", "label": "Desempeño mensual"},
                {"key": "recovery", "label": "Recuperación de inversión"},
                {"key": "comparison", "label": "Comparativo real vs proyectado"},
                {"key": "debt", "label": "Estado de deuda"},
                {"key": "roi", "label": "ROI / payback / VAN / TIR"},
            ],
        }
    )

    if export_format == "xlsx":
        return _export_project_xlsx(context, report_key)
    if export_format == "pdf":
        return _export_project_pdf(context, report_key)
    if export_format == "csv" and report_key == "investment":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="proyecto_inversion_{project.pk}_investment.csv"'
        writer = csv.writer(response)
        writer.writerow(["Fecha", "Categoría", "Descripción", "Monto total", "Referencia compra"])
        for row in context["expenses"]:
            writer.writerow([row.fecha.isoformat(), row.get_categoria_display(), row.descripcion, row.monto_total, row.referencia_compra])
        return response

    return render(request, "reportes/proyecto_inversion_detail.html", context)


@login_required
def proyecto_bamoa_wizard(request: HttpRequest) -> HttpResponse:
    """
    Wizard multi-paso para crear Bamoa 2026 con partidas de inversión.
    """
    from django.db.models import Count, Sum
    from reportes.services_investment_projects import _benchmark_sucursales_activas

    _require_reportes_access(request.user)

    benchmark: dict[str, object] = {}
    bm_error = None
    try:
        benchmark = _benchmark_sucursales_activas(meses=12)
    except Exception as exc:
        bm_error = str(exc)

    guamuchil_por_categoria: dict[str, dict[str, object]] = {}
    guamuchil_total = Decimal("0")
    partidas_referencia_gml: list[dict[str, object]] = []
    try:
        p_gml = ProyectoInversion.objects.get(pk=1)
        cats = (
            p_gml.gastos_inversion.values("categoria")
            .annotate(total=Sum("monto_total"), num_partidas=Count("id"))
            .order_by("categoria")
        )
        category_labels = dict(ProyectoInversionGasto.CATEGORIA_CHOICES)
        for category in cats:
            guamuchil_por_categoria[category["categoria"]] = {
                "total": float(category["total"] or 0),
                "num_partidas": category["num_partidas"],
                "label": category_labels.get(category["categoria"], category["categoria"]),
            }
        guamuchil_total = _parse_decimal(str(p_gml.monto_inversion_real))
        partidas_referencia_gml = list(
            p_gml.gastos_inversion.values("categoria", "descripcion", "proveedor_nombre", "monto_total")
            .order_by("categoria", "-monto_total")[:30]
        )
        for partida in partidas_referencia_gml:
            partida["monto_total"] = float(partida["monto_total"] or 0)
            partida["categoria_label"] = category_labels.get(partida["categoria"], partida["categoria"])
    except ProyectoInversion.DoesNotExist:
        pass
    except Exception as exc:
        import logging

        logging.getLogger(__name__).warning("guamuchil benchmark error: %s", exc)

    if request.method == "POST" and request.POST.get("action") == "create_bamoa_project":
        errores = []
        nombre_proyecto = (request.POST.get("nombre_proyecto") or "Apertura Bamoa 2026").strip()
        fecha_inicio = _parse_date(request.POST.get("fecha_inicio"))
        fecha_apertura = _parse_date(request.POST.get("fecha_apertura"))
        ventas_base = _parse_decimal(request.POST.get("ventas_promedio_base"))

        try:
            partidas_raw = json.loads(request.POST.get("partidas_json") or "[]")
        except (TypeError, ValueError, json.JSONDecodeError):
            partidas_raw = []
            errores.append("Error al leer las partidas de inversión.")

        allowed_categories = {key for key, _ in ProyectoInversionGasto.CATEGORIA_CHOICES}
        partidas_validas = []
        for partida in partidas_raw:
            if not isinstance(partida, dict):
                continue
            descripcion = (partida.get("descripcion") or "").strip()
            monto = _parse_decimal(str(partida.get("monto", 0)))
            iva = _parse_decimal(str(partida.get("iva", 0)))
            if not descripcion or monto <= 0:
                continue
            categoria = (partida.get("categoria") or ProyectoInversionGasto.CATEGORIA_OTROS).strip()
            if categoria not in allowed_categories:
                categoria = ProyectoInversionGasto.CATEGORIA_OTROS
            partidas_validas.append(
                {
                    "categoria": categoria,
                    "subcategoria": (partida.get("subcategoria") or "").strip(),
                    "descripcion": descripcion,
                    "proveedor_nombre": (partida.get("proveedor_nombre") or "").strip(),
                    "monto": monto,
                    "iva": iva,
                    "monto_total": monto + iva,
                    "notas": (partida.get("notas") or "").strip(),
                }
            )

        deposito_renta = _parse_decimal(request.POST.get("deposito_renta"), default="0")
        if deposito_renta > 0:
            partidas_validas.append(
                {
                    "categoria": ProyectoInversionGasto.CATEGORIA_OTROS,
                    "subcategoria": "Depósito",
                    "descripcion": "Depósito de renta Bamoa",
                    "proveedor_nombre": "",
                    "monto": deposito_renta,
                    "iva": Decimal("0"),
                    "monto_total": deposito_renta,
                    "notas": "Capturado desde supuestos operativos del wizard.",
                }
            )

        if not fecha_inicio:
            errores.append("La fecha de inicio es obligatoria.")
        if not partidas_validas:
            errores.append("Debes capturar al menos una partida de inversión con monto mayor a cero.")
        if ventas_base <= 0:
            errores.append("Las ventas promedio base deben ser mayores a cero.")
        existing_project = ProyectoInversion.objects.filter(nombre_proyecto=nombre_proyecto).first()
        if existing_project:
            messages.warning(request, f"Ya existe el proyecto '{nombre_proyecto}'. Se abrió el registro existente.")
            return redirect("reportes:proyecto_inversion_detail", project_id=existing_project.pk)
        if errores:
            for error in errores:
                messages.error(request, error)
            return redirect("reportes:proyecto_bamoa_wizard")

        inversion_total = sum((partida["monto_total"] for partida in partidas_validas), Decimal("0"))
        renta = _parse_decimal(request.POST.get("renta_mensual"), default="4000")
        nomina = _parse_decimal(request.POST.get("nomina_mensual"))
        servicios = _parse_decimal(request.POST.get("servicios_mensual"))
        marketing = _parse_decimal(request.POST.get("marketing_mensual"))
        otros_fijos = _parse_decimal(request.POST.get("otros_fijos_mensual"))
        gastos_totales = renta + nomina + servicios + marketing + otros_fijos
        margen_base = _parse_decimal(request.POST.get("margen_bruto_pct"), default="45")
        crecimiento = _parse_decimal(request.POST.get("crecimiento_mensual_pct"), default="0.8")
        horizonte_meses = _parse_int(request.POST.get("horizonte_meses"), default=36) or 36

        with transaction.atomic():
            project = ProyectoInversion.objects.create(
                nombre_proyecto=nombre_proyecto,
                tipo_proyecto=ProyectoInversion.TIPO_APERTURA_SUCURSAL,
                sucursal_relacionada_id=_parse_int(request.POST.get("sucursal_relacionada_id")) or None,
                fecha_inicio=fecha_inicio,
                fecha_apertura=fecha_apertura,
                responsable=request.user,
                estatus=ProyectoInversion.ESTATUS_PLANEACION,
                monto_inversion_planeado=inversion_total,
                capital_inicial_aportado=inversion_total,
                deuda_asociada=_parse_decimal(request.POST.get("deuda_asociada")),
                tasa_interes_anual=_parse_decimal(request.POST.get("tasa_interes_anual")),
                plazo_deuda_meses=_parse_int(request.POST.get("plazo_deuda_meses")),
                pago_mensual_deuda_estimado=_parse_decimal(request.POST.get("pago_mensual_deuda_estimado")),
                discount_rate=_parse_decimal(request.POST.get("discount_rate"), default="12"),
                roi_objetivo=_parse_decimal(request.POST.get("roi_objetivo"), default="25"),
                payback_objetivo_meses=_parse_int(request.POST.get("payback_objetivo_meses"), default=24) or 24,
                recovery_strategy=ProyectoInversion.RECOVERY_FULL_NET_CASHFLOW,
                recovery_percentage=Decimal("1"),
                cierre_por_recuperacion_total=True,
                observaciones=(request.POST.get("observaciones") or "").strip(),
                metadata={
                    "benchmark_snapshot": benchmark,
                    "guamuchil_referencia": guamuchil_por_categoria,
                    "supuestos_operativos": {
                        "renta": float(renta),
                        "deposito_renta": float(deposito_renta),
                        "nomina": float(nomina),
                        "servicios": float(servicios),
                        "marketing": float(marketing),
                        "otros_fijos": float(otros_fijos),
                        "gastos_totales": float(gastos_totales),
                    },
                    "creado_desde": "proyecto_bamoa_wizard_v2",
                },
            )

            for partida in partidas_validas:
                ProyectoInversionGasto.objects.create(
                    proyecto=project,
                    fecha=fecha_inicio,
                    categoria=partida["categoria"],
                    subcategoria=partida["subcategoria"],
                    descripcion=partida["descripcion"],
                    proveedor_nombre=partida["proveedor_nombre"],
                    monto=partida["monto"],
                    iva=partida["iva"],
                    monto_total=partida["monto_total"],
                    metodo_pago="",
                    financiado=False,
                    referencia_contable=f"BAMOA_PLAN_{project.pk}",
                    notas=partida["notas"],
                    capturado_por=request.user,
                )

            escenarios_def = [
                {
                    "nombre": "Conservador",
                    "tipo_escenario": ProyectoInversionEscenario.TIPO_CONSERVADOR,
                    "ventas_promedio_mensuales": (ventas_base * Decimal("0.88")).quantize(Decimal("0.01")),
                    "crecimiento_mensual_pct": (crecimiento * Decimal("0.85")).quantize(Decimal("0.0001")),
                    "margen_bruto_pct": (margen_base * Decimal("0.96")).quantize(Decimal("0.0001")),
                    "gastos_operativos_mensuales": (gastos_totales * Decimal("1.06")).quantize(Decimal("0.01")),
                    "notas": "Conservador: -12% ventas, +6% gastos contra base.",
                },
                {
                    "nombre": "Base",
                    "tipo_escenario": ProyectoInversionEscenario.TIPO_BASE,
                    "ventas_promedio_mensuales": ventas_base,
                    "crecimiento_mensual_pct": crecimiento,
                    "margen_bruto_pct": margen_base,
                    "gastos_operativos_mensuales": gastos_totales,
                    "notas": "Base: supuestos capturados en wizard.",
                },
                {
                    "nombre": "Optimista",
                    "tipo_escenario": ProyectoInversionEscenario.TIPO_OPTIMISTA,
                    "ventas_promedio_mensuales": (ventas_base * Decimal("1.12")).quantize(Decimal("0.01")),
                    "crecimiento_mensual_pct": (crecimiento * Decimal("1.15")).quantize(Decimal("0.0001")),
                    "margen_bruto_pct": (margen_base * Decimal("1.04")).quantize(Decimal("0.0001")),
                    "gastos_operativos_mensuales": (gastos_totales * Decimal("0.97")).quantize(Decimal("0.01")),
                    "notas": "Optimista: +12% ventas, -3% gastos contra base.",
                },
            ]

            scenario_service = ProyectoInversionScenarioService()
            for scenario_data in escenarios_def:
                scenario = ProyectoInversionEscenario.objects.create(
                    proyecto=project,
                    horizonte_meses=horizonte_meses,
                    capturado_por=request.user,
                    **scenario_data,
                )
                scenario_service.compute(project, scenario)

            try:
                ProyectoInversionRefreshService().refresh_project(project, user=request.user)
            except Exception as exc:
                import logging

                logging.getLogger(__name__).warning("refresh inicial Bamoa: %s", exc)

            log_event(
                request.user,
                "CREATE",
                "reportes.ProyectoInversion",
                project.pk,
                payload={
                    "nombre_proyecto": project.nombre_proyecto,
                    "origen": "proyecto_bamoa_wizard_v2",
                    "partidas": len(partidas_validas),
                    "inversion_total": float(inversion_total),
                    "escenarios_creados": 3,
                },
            )
            messages.success(
                request,
                f"Proyecto '{project.nombre_proyecto}' creado con {len(partidas_validas)} partidas y 3 escenarios.",
            )
            return redirect("reportes:proyecto_inversion_detail", project_id=project.pk)

    categorias_choices = ProyectoInversionGasto.CATEGORIA_CHOICES
    categorias_json = json.dumps([{"value": value, "label": label} for value, label in categorias_choices])

    return render(
        request,
        "reportes/proyecto_bamoa_wizard.html",
        {
            "module_tabs": _project_module_tabs("sucursales"),
            "benchmark": benchmark,
            "bm_error": bm_error,
            "guamuchil_por_categoria": guamuchil_por_categoria,
            "guamuchil_total": float(guamuchil_total),
            "guamuchil_por_categoria_json": json.dumps(guamuchil_por_categoria),
            "partidas_referencia_gml_json": json.dumps(partidas_referencia_gml),
            "categorias_choices": categorias_choices,
            "categorias_json": categorias_json,
            "ventas_sugeridas": float(benchmark.get("ventas_mensuales_avg") or 0),
            "gastos_sugeridos": float(benchmark.get("gastos_operativos_avg") or 0),
            "ticket_sugerido": float(benchmark.get("ticket_promedio") or 0),
            "sucursales": Sucursal.objects.filter(activa=True).order_by("nombre"),
            "renta_sugerida": 4000,
            "deposito_sugerido": 4000,
            "today": timezone.localdate(),
        },
    )


@login_required
def api_bamoa_guamuchil_benchmark(request: HttpRequest) -> JsonResponse:
    """Retorna benchmark de inversión Guamúchil para el wizard Bamoa."""
    from django.db.models import Count, Sum

    _require_reportes_access(request.user)
    try:
        project = ProyectoInversion.objects.get(pk=1)
        category_labels = dict(ProyectoInversionGasto.CATEGORIA_CHOICES)
        categories = list(
            project.gastos_inversion.values("categoria")
            .annotate(total=Sum("monto_total"), num_partidas=Count("id"))
            .order_by("categoria")
        )
        for category in categories:
            category["total"] = float(category["total"] or 0)
            category["label"] = category_labels.get(category["categoria"], category["categoria"])
        return JsonResponse(
            {
                "ok": True,
                "categorias": categories,
                "total": float(project.monto_inversion_real),
                "partidas": project.gastos_inversion.count(),
            }
        )
    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=500)


@login_required
def proyecto_viabilidad_export_excel(request: HttpRequest, project_id: int) -> HttpResponse:
    """Exporta el análisis de viabilidad del proyecto a Excel."""
    _require_reportes_access(request.user)
    project = get_object_or_404(ProyectoInversion, pk=project_id)
    escenarios = list(project.escenarios.all().order_by("tipo_escenario", "nombre"))
    snapshots = list(
        project.snapshots_mensuales.filter(
            data_source__in=[
                ProyectoInversionSnapshotMensual.DATA_SOURCE_FACT,
                ProyectoInversionSnapshotMensual.DATA_SOURCE_FALLBACK,
                ProyectoInversionSnapshotMensual.DATA_SOURCE_ESTIMATED,
            ]
        ).order_by("periodo")
    )

    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    for label, value in [
        ("Proyecto", project.nombre_proyecto),
        ("Tipo", project.get_tipo_proyecto_display()),
        ("Estatus", project.get_estatus_display()),
        ("Fecha inicio", str(project.fecha_inicio)),
        ("Fecha apertura", str(project.fecha_apertura or "")),
        ("Inversión planeada", project.monto_inversion_planeado),
        ("Inversión real", project.monto_inversion_real),
        ("Capital aportado", project.capital_inicial_aportado),
        ("Deuda asociada", project.deuda_asociada),
        ("Tasa interés anual", project.tasa_interes_anual),
        ("Plazo deuda meses", project.plazo_deuda_meses),
        ("ROI objetivo", project.roi_objetivo),
        ("Payback objetivo meses", project.payback_objetivo_meses),
        ("Discount rate", project.discount_rate),
        ("Observaciones", project.observaciones),
    ]:
        ws_resumen.append([label, value])

    benchmark = (project.metadata or {}).get("benchmark_snapshot", {})
    ws_benchmark = wb.create_sheet("Benchmark")
    ws_benchmark.append(["Métrica", "Valor"])
    for key in [
        "data_source",
        "periodo_inicio",
        "periodo_fin",
        "ticket_promedio",
        "ventas_mensuales_avg",
        "ventas_diarias_avg",
        "margen_bruto_pct_avg",
        "gastos_operativos_avg",
        "utilidad_operativa_avg",
        "payback_real_avg_meses",
        "roi_acumulado_avg",
    ]:
        ws_benchmark.append([key, benchmark.get(key, "")])

    ws_escenarios = wb.create_sheet("Escenarios")
    ws_escenarios.append(
        [
            "Nombre",
            "Tipo",
            "Ventas promedio mensual",
            "Crecimiento mensual %",
            "Margen bruto %",
            "Gastos operativos",
            "Horizonte meses",
            "Payback meses",
            "ROI anual %",
            "Flujo libre mensual",
            "Notas",
        ]
    )
    for escenario in escenarios:
        resultados = escenario.resultados or {}
        ws_escenarios.append(
            [
                escenario.nombre,
                escenario.get_tipo_escenario_display(),
                escenario.ventas_promedio_mensuales,
                escenario.crecimiento_mensual_pct,
                escenario.margen_bruto_pct,
                escenario.gastos_operativos_mensuales,
                escenario.horizonte_meses,
                resultados.get("payback_months", ""),
                resultados.get("annual_roi_pct", ""),
                resultados.get("monthly_free_cashflow", ""),
                escenario.notas,
            ]
        )

    ws_snapshots = wb.create_sheet("Snapshots mensuales")
    ws_snapshots.append(
        [
            "Periodo",
            "Ventas",
            "Utilidad bruta",
            "Gastos operativos",
            "Utilidad operativa",
            "Flujo libre",
            "Recuperación acumulada",
            "Saldo pendiente",
            "ROI acumulado %",
            "Payback real meses",
            "Health",
            "Data source",
        ]
    )
    for snapshot in snapshots:
        ws_snapshots.append(
            [
                str(snapshot.periodo),
                snapshot.ventas_mensuales or 0,
                snapshot.utilidad_bruta or 0,
                snapshot.gastos_operativos or 0,
                snapshot.utilidad_operativa or 0,
                snapshot.flujo_libre or 0,
                snapshot.recuperacion_acumulada or 0,
                snapshot.saldo_pendiente or 0,
                snapshot.roi_acumulado or 0,
                snapshot.payback_real_meses or 0,
                snapshot.health_status,
                snapshot.data_source,
            ]
        )

    filename = f"viabilidad_{project.nombre_proyecto.replace(' ', '_')}_{timezone.localdate()}.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
