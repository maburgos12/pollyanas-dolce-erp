import csv
import re
from collections import defaultdict
from io import BytesIO
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from types import SimpleNamespace
from urllib.parse import urlencode
from calendar import monthrange
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import Workbook
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.core.exceptions import PermissionDenied
from django.db.models import Sum, Max, Count, Q
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from unidecode import unidecode

from core.access import (
    can_build_product_closure,
    can_manage_orquestacion,
    can_lock_product_closure,
    can_rebuild_product_closure,
    can_view_product_closure,
    can_view_reportes,
)
from core.audit import log_event
from core.cache_versions import get_or_set_versioned_cache
from core.branch_catalog import eligible_operational_branch_qs
from core.models import AuditLog
from inventario.models import ExistenciaInsumo, MovimientoInventario
from control.models import MermaPOS, VentaPOS
from maestros.models import CostoInsumo, Insumo
from maestros.utils.canonical_catalog import canonicalized_active_insumos, enterprise_readiness_profile, latest_costo_canonico
from recetas.models import (
    ProductoMonthClosure,
    Receta,
    LineaReceta,
    RecetaCostoSemanal,
    VentaHistorica,
    PlanProduccion,
    PlanProduccionItem,
    PronosticoVenta,
    SolicitudVenta,
)
from compras.models import SolicitudCompra, OrdenCompra, RecepcionCompra
from pos_bridge.models import (
    PointBranch,
    PointDailyBranchIndicator,
    PointExtractionLog,
    PointSalesDailyCategoryFact,
    PointSalesDailyProductFact,
    PointSyncJob,
)
from pos_bridge.config import load_point_bridge_settings
from pos_bridge.tasks.celery_tasks import task_operations_automation_cycle, task_visible_cut_refresh_cycle
from pos_bridge.services.product_month_closure_service import ProductMonthClosureError, ProductMonthClosureService
from ventas.services.sales_canonical_source import (
    OFFICIAL_POINT_SOURCE as CANONICAL_OFFICIAL_POINT_SOURCE,
    POINT_BRIDGE_SALES_SOURCE as CANONICAL_POINT_BRIDGE_SALES_SOURCE,
    RECENT_POINT_SOURCE as CANONICAL_RECENT_POINT_SOURCE,
    authoritative_sales_max_date as canonical_authoritative_sales_max_date,
    build_sales_source_context as build_canonical_sales_source_context,
    canonical_point_max_date as get_canonical_point_max_date,
    canonical_point_previous_dates as get_canonical_point_previous_dates,
    get_sales_source_context as get_canonical_sales_source_context,
    official_sales_stage_max_date as get_official_sales_stage_max_date,
    operational_sales_filters as canonical_operational_sales_filters,
    operational_sales_rows_for_date as canonical_operational_sales_rows_for_date,
    point_sales_month_total as canonical_point_sales_month_total,
    recent_sales_stage_max_date as get_recent_sales_stage_max_date,
    sales_history_queryset as canonical_sales_history_queryset,
    sales_previous_dates as canonical_sales_previous_dates,
    sales_rows_for_date as canonical_sales_rows_for_date,
    sales_rows_for_month as canonical_sales_rows_for_month,
    v2_category_sales_max_date as canonical_v2_category_sales_max_date,
    v2_product_sales_max_date as canonical_v2_product_sales_max_date,
)
from ventas.services.sales_read_service import get_daily_sales_bulk, get_sales_range
from ventas.models import VentaAutoritativaPoint
from recetas.utils.derived_product_presentations import get_total_cost_map

from .bi_utils import compute_bi_snapshot
from .auto_production_service import (
    approve_production_order,
    execute_production_order,
    generate_daily_production_orders,
    release_production_order,
    sync_production_execution_logs,
)
from .auto_purchase_service import generate_purchase_requests_from_production, list_auto_purchase_snapshots
from .alert_service import generate_operational_alerts, resolve_alert
from .dashboard_sales_dataset import get_dashboard_sales_dataset
from .daily_operational_closure_service import build_daily_operational_closure
from .forecast_service import build_daily_forecast_context
from .production_projection_supply_service import build_projection_supply_context
from .production_supply_service import build_production_supply_context
from .executive_panels import (
    _active_sales_queryset,
    _recipe_cost_map_for_sales_lens,
    build_executive_bi_panels,
    build_monthly_yoy_panel,
    build_profitability_panel,
    build_sales_forecast_panel,
)
from .models import (
    Alert,
    CargaGastoOperativoArchivo,
    CorteOficialDiario,
    OperationsMetricSnapshot,
    PresupuestoImport,
    ProductionOrder,
)
from .services_budget_area_upload import BudgetAreaUploadService
from .operations_metrics_service import rebuild_operations_metrics
from .services_operating_expense_automation import OperatingExpenseImportAutomationService

POINT_BRIDGE_SALES_SOURCE = CANONICAL_POINT_BRIDGE_SALES_SOURCE
OFFICIAL_POINT_SOURCE = CANONICAL_OFFICIAL_POINT_SOURCE
RECENT_POINT_SOURCE = CANONICAL_RECENT_POINT_SOURCE
DEFAULT_DAILY_SALES_EXCLUDED_BRANCHES = {
    "CEDIS",
    "ALMACEN",
    "PRODUCCION_CRUCERO",
    "DEVOLUCIONES",
    "MATRIZDBG",
}
VALID_DAILY_SALES_ZERO_EXCEPTIONS: dict[date, dict[str, str]] = {
    date(2026, 1, 7): {"LEYVA": "Cierre por remodelación / incidencia operativa validada."},
    date(2026, 1, 8): {"LEYVA": "Cierre por remodelación / incidencia operativa validada."},
    date(2026, 1, 9): {"LEYVA": "Cierre por remodelación / incidencia operativa validada."},
    date(2026, 1, 10): {"LEYVA": "Cierre por remodelación / incidencia operativa validada."},
    date(2026, 1, 11): {"LEYVA": "Cierre por remodelación / incidencia operativa validada."},
    date(2026, 1, 12): {"LEYVA": "Cierre por remodelación / incidencia operativa validada."},
    date(2026, 1, 13): {"LEYVA": "Cierre por remodelación / incidencia operativa validada."},
    date(2026, 2, 3): {"EL_TUNEL": "Sucursal sin apertura validada en martes."},
    date(2026, 2, 10): {"EL_TUNEL": "Sucursal sin apertura validada en martes."},
    date(2026, 2, 17): {"EL_TUNEL": "Sucursal sin apertura validada en martes."},
    date(2026, 2, 24): {"EL_TUNEL": "Sucursal sin apertura validada en martes."},
    date(2026, 3, 3): {"EL_TUNEL": "Sucursal sin apertura validada en martes."},
    date(2026, 3, 10): {"EL_TUNEL": "Sucursal sin apertura validada en martes."},
    date(2026, 3, 17): {"EL_TUNEL": "Sucursal sin apertura validada en martes."},
    date(2026, 3, 24): {"EL_TUNEL": "Sucursal sin apertura validada en martes."},
    date(2026, 4, 3): {
        "COLOSIO": "Cierre operativo validado; solo abrió Matriz.",
        "CRUCERO": "Cierre operativo validado; solo abrió Matriz.",
        "EL_TUNEL": "Cierre operativo validado; solo abrió Matriz.",
        "GUAMUCHIL": "Cierre operativo validado; solo abrió Matriz.",
        "LAS_GLORIAS": "Cierre operativo validado; solo abrió Matriz.",
        "LEYVA": "Cierre operativo validado; solo abrió Matriz.",
        "PAYAN": "Cierre operativo validado; solo abrió Matriz.",
        "PLAZA_NIO": "Cierre operativo validado; solo abrió Matriz.",
    },
    date(2026, 4, 4): {"MATRIZ": "Cierre operativo validado; abrió el resto de sucursales."},
    date(2026, 4, 7): {"LEYVA": "Incidencia de red validada."},
}
BI_FORCE_REFRESH_LOCK_KEY = "reportes:bi-force-refresh-lock"
BI_FORCE_REFRESH_LOCK_SECONDS = 15 * 60
INTEGRATIONS_OPERATIONAL_REFRESH_LOCK_KEY = "integraciones:operational-refresh-lock"


def _reportes_module_tabs(active: str) -> list[dict[str, str | bool]]:
    tabs = [
        ("ventas", reverse("reportes:ventas"), "Ventas"),
        ("cierre_operativo", reverse("reportes:cierre_operativo"), "Cierre diario"),
        ("cierre_producto", reverse("reportes:cierre_producto"), "Cierre producto"),
        ("financiero", reverse("reportes:financiero"), "Financiero"),
        ("presupuestos", reverse("reportes:presupuesto_importar_por_area"), "Presupuestos"),
        ("gastos_operativos", reverse("reportes:gastos_operativos_importar"), "Importar gastos"),
        ("consumo", reverse("reportes:consumo"), "Consumo"),
        ("faltantes", reverse("reportes:faltantes"), "Faltantes"),
        ("bi", reverse("reportes:bi"), "BI"),
        ("proyectos", reverse("reportes:proyectos_inversion"), "Proyectos inversión"),
    ]
    return [
        {"key": key, "url": url, "label": label, "active": key == active}
        for key, url, label in tabs
    ]


def _bi_force_refresh_redirect_target(request: HttpRequest) -> str:
    next_url = (request.POST.get("next") or request.GET.get("next") or "").strip()
    if next_url.startswith("/") and not next_url.startswith("//"):
        return next_url
    return reverse("reportes:bi")


def _queue_operations_refresh(
    *,
    reference_date: date,
    lookback_days: int,
    triggered_by,
) -> dict[str, object]:
    payload = {
        "reference_date": reference_date.isoformat(),
        "lookback_days": int(lookback_days or 7),
        "sucursal_id": None,
        "skip_refresh": False,
        "triggered_by_id": getattr(triggered_by, "id", None),
        "trigger": "reportes_bi_ui_inline",
    }
    try:
        async_result = task_operations_automation_cycle.delay(
            reference_date_iso=reference_date.isoformat(),
            lookback_days=int(lookback_days or 7),
            sucursal_id=None,
            skip_refresh=False,
            triggered_by_id=getattr(triggered_by, "id", None),
        )
    except Exception as exc:
        log_event(
            triggered_by,
            "INTEGRATIONS_OPERATIONAL_REFRESH_FAILED",
            "reportes.AnalyticRefreshWindow",
            reference_date.isoformat(),
            payload={**payload, "error": str(exc)},
        )
        cache.delete(BI_FORCE_REFRESH_LOCK_KEY)
        cache.delete(INTEGRATIONS_OPERATIONAL_REFRESH_LOCK_KEY)
        raise
    payload["task_id"] = str(getattr(async_result, "id", "") or "")
    return payload


def _queue_visible_cut_refresh(
    *,
    reference_date: date,
    triggered_by,
) -> dict[str, object]:
    payload = {
        "reference_date": reference_date.isoformat(),
        "lookback_days": 1,
        "lag_days": 0,
        "scope": "visible_cut",
        "triggered_by_id": getattr(triggered_by, "id", None),
        "trigger": "dashboard_visible_cut_refresh",
    }
    try:
        async_result = task_visible_cut_refresh_cycle.delay(
            reference_date_iso=reference_date.isoformat(),
            triggered_by_id=getattr(triggered_by, "id", None),
        )
    except Exception as exc:
        log_event(
            triggered_by,
            "INTEGRATIONS_OPERATIONAL_REFRESH_FAILED",
            "reportes.AnalyticRefreshWindow",
            reference_date.isoformat(),
            payload={**payload, "error": str(exc)},
        )
        cache.delete(BI_FORCE_REFRESH_LOCK_KEY)
        cache.delete(INTEGRATIONS_OPERATIONAL_REFRESH_LOCK_KEY)
        raise
    payload["task_id"] = str(getattr(async_result, "id", "") or "")
    return payload


@login_required
def bi_force_refresh(request: HttpRequest) -> HttpResponse:
    if request.method != "POST":
        raise PermissionDenied
    if not can_manage_orquestacion(request.user):
        raise PermissionDenied

    redirect_url = _bi_force_refresh_redirect_target(request)
    refresh_scope = (request.POST.get("refresh_scope") or "").strip().lower()
    reference_date_raw = (request.POST.get("reference_date") or "").strip()
    try:
        reference_date = date.fromisoformat(reference_date_raw) if reference_date_raw else timezone.localdate()
    except ValueError:
        reference_date = timezone.localdate()
    try:
        lookback_days = max(1, min(int(request.POST.get("lookback_days") or 7), 30))
    except (TypeError, ValueError):
        lookback_days = 7
    if refresh_scope == "cutoff":
        lookback_days = 1

    if not cache.add(BI_FORCE_REFRESH_LOCK_KEY, reference_date.isoformat(), BI_FORCE_REFRESH_LOCK_SECONDS):
        messages.warning(
            request,
            "El corte ya se está actualizando. En unos momentos verás el dato nuevo en el tablero.",
        )
        return redirect(redirect_url)

    try:
        log_event(
            request.user,
            "REPORTES_BI_FORCE_REFRESH_REQUESTED",
            "reportes.AnalyticRefreshWindow",
            reference_date.isoformat(),
            payload={
                "reference_date": reference_date.isoformat(),
                "lookback_days": lookback_days,
                "scope": "visible_cut" if refresh_scope == "cutoff" else "analytics_and_operations_cycle",
                "trigger": "reportes_bi_ui",
            },
        )
        if refresh_scope == "cutoff":
            _queue_visible_cut_refresh(
                reference_date=reference_date,
                triggered_by=request.user,
            )
        else:
            _queue_operations_refresh(
                reference_date=reference_date,
                lookback_days=lookback_days,
                triggered_by=request.user,
            )
    except Exception as exc:
        messages.error(
            request,
            (
                "La actualización operativa no se pudo completar. "
                f"Referencia {reference_date.isoformat()}: {exc}"
            ),
        )
        return redirect(redirect_url)

    messages.success(
        request,
        (
            "Actualización del corte en proceso. "
            f"Referencia {reference_date.isoformat()}."
        ),
    )
    return redirect(redirect_url)


def _bi_cached_value(
    *,
    runtime_cache: dict[str, object],
    section: str,
    builder,
    parts: tuple[object, ...] = (),
):
    return get_or_set_versioned_cache(
        key_parts=("erp", "bi", section, *parts),
        scopes=("dashboard",),
        builder=builder,
        runtime_cache=runtime_cache,
    )


def _to_decimal(value, default: str = "0") -> Decimal:
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)


def _parse_ui_date(value: str | None) -> date:
    if not value:
        return timezone.localdate()
    try:
        return date.fromisoformat(str(value))
    except ValueError:
        return timezone.localdate()


def _normalize_branch_token(value: str | None) -> str:
    normalized = unidecode((value or "").strip()).upper()
    normalized = re.sub(r"[^A-Z0-9]+", "_", normalized)
    return normalized.strip("_")


def _daily_sales_excluded_branch_tokens() -> set[str]:
    settings = load_point_bridge_settings()
    configured = {
        _normalize_branch_token(item)
        for item in (settings.sales_excluded_branches or [])
        if _normalize_branch_token(item)
    }
    return configured or set(DEFAULT_DAILY_SALES_EXCLUDED_BRANCHES)


def _is_excluded_daily_sales_branch(*, code: str | None, name: str | None) -> bool:
    excluded = _daily_sales_excluded_branch_tokens()
    return _normalize_branch_token(code) in excluded or _normalize_branch_token(name) in excluded


def _required_daily_sales_branches(target_date: date) -> list[dict[str, object]]:
    required_by_key: dict[str, dict[str, object]] = {}
    eligible_branches = list(eligible_operational_branch_qs(target_date))
    eligible_branch_ids = {branch.id for branch in eligible_branches}

    def register_branch(*, branch_id: int, branch_code: str | None, branch_name: str | None, source_label: str) -> None:
        dedupe_key = _normalize_branch_token(branch_name) or _normalize_branch_token(branch_code) or str(branch_id)
        current = required_by_key.get(dedupe_key)
        candidate = {
            "branch_id": branch_id,
            "branch_code": branch_code,
            "branch_name": branch_name,
            "source": source_label,
        }
        if current is None:
            required_by_key[dedupe_key] = candidate
            return
        # If there are duplicated ERP branches for the same operating location, prefer the Point-mapped one.
        if source_label == "point_activa" and current.get("source") != "point_activa":
            required_by_key[dedupe_key] = candidate

    for branch in eligible_branches:
        if _is_excluded_daily_sales_branch(code=branch.codigo, name=branch.nombre):
            continue
        register_branch(
            branch_id=branch.id,
            branch_code=branch.codigo,
            branch_name=branch.nombre,
            source_label="erp_operativa",
        )

    point_mapped_rows = (
        PointBranch.objects.filter(erp_branch__isnull=False, status=PointBranch.STATUS_ACTIVE)
        .select_related("erp_branch")
        .order_by("erp_branch_id", "id")
    )
    for point_branch in point_mapped_rows:
        erp_branch = point_branch.erp_branch
        if erp_branch is None:
            continue
        if erp_branch.id not in eligible_branch_ids:
            continue
        if erp_branch.fecha_apertura and erp_branch.fecha_apertura > target_date:
            continue
        if _is_excluded_daily_sales_branch(code=erp_branch.codigo, name=erp_branch.nombre):
            continue
        register_branch(
            branch_id=erp_branch.id,
            branch_code=erp_branch.codigo,
            branch_name=erp_branch.nombre,
            source_label="point_activa",
        )

    return sorted(
        required_by_key.values(),
        key=lambda item: (str(item.get("branch_code") or ""), str(item.get("branch_name") or "")),
    )


def _indicator_present_daily_sales_branch_ids(target_date: date) -> set[int]:
    required_branch_ids = {int(branch["branch_id"]) for branch in _required_daily_sales_branches(target_date)}
    indicator_ids = set(
        PointDailyBranchIndicator.objects.filter(indicator_date=target_date)
        .exclude(branch__erp_branch_id__isnull=True)
        .values_list("branch__erp_branch_id", flat=True)
    )
    return {
        int(branch_id)
        for branch_id in indicator_ids
        if branch_id and int(branch_id) in required_branch_ids
    }


def _official_backfill_present_daily_sales_branch_ids(target_date: date) -> set[int]:
    required_branches = _required_daily_sales_branches(target_date)
    branch_ids_by_external = {
        str(branch.external_id): int(branch.erp_branch_id)
        for branch in PointBranch.objects.filter(
            erp_branch_id__in=[int(item["branch_id"]) for item in required_branches]
        ).exclude(erp_branch_id__isnull=True)
        if branch.external_id
    }
    if not branch_ids_by_external:
        return set()

    log_contexts = PointExtractionLog.objects.filter(
        sync_job__job_type=PointSyncJob.JOB_TYPE_SALES,
        level=PointExtractionLog.LEVEL_INFO,
        message__startswith="Backfill oficial ",
        context__sale_date=target_date.isoformat(),
        context__has_key="rows_imported",
    ).values_list("context", flat=True)
    present_branch_ids: set[int] = set()
    for context in log_contexts:
        payload = context or {}
        branch_external_id = str(payload.get("branch_external_id") or "").strip()
        erp_branch_id = branch_ids_by_external.get(branch_external_id)
        if erp_branch_id is not None:
            present_branch_ids.add(int(erp_branch_id))
    return present_branch_ids


def _validated_zero_daily_sales_branch_codes(target_date: date) -> set[str]:
    return set(VALID_DAILY_SALES_ZERO_EXCEPTIONS.get(target_date, {}).keys())


def _present_daily_sales_branch_ids(*, source: dict[str, object], target_date: date) -> set[int]:
    indicator_ids = _indicator_present_daily_sales_branch_ids(target_date)
    official_backfill_ids = _official_backfill_present_daily_sales_branch_ids(target_date)
    if source["mode"] == "point_stage":
        sales_ids = set(
            _operational_sales_rows_for_date(target_date)
            .exclude(branch__erp_branch_id__isnull=True)
            .values_list("branch__erp_branch_id", flat=True)
        )
        indicator_ids = set(
            PointDailyBranchIndicator.objects.filter(indicator_date=target_date)
            .exclude(branch__erp_branch_id__isnull=True)
            .values_list("branch__erp_branch_id", flat=True)
        )
        return {int(branch_id) for branch_id in sales_ids.union(indicator_ids).union(official_backfill_ids) if branch_id}
    if source["mode"] in {"point_history", "historical_fallback"}:
        sales_ids = {
            int(branch_id)
            for branch_id in _sales_rows_for_date(source, target_date)
            .exclude(sucursal_id__isnull=True)
            .values_list("sucursal_id", flat=True)
            if branch_id
        }
        return sales_ids.union(indicator_ids).union(official_backfill_ids)
    return indicator_ids.union(official_backfill_ids)


def _canonical_catalog_maps(limit: int = 2000) -> tuple[dict[int, dict], dict[int, dict]]:
    canonical_rows = canonicalized_active_insumos(limit=limit)
    member_to_row = {}
    canonical_by_id = {}
    for row in canonical_rows:
        canonical = row["canonical"]
        canonical_by_id[canonical.id] = row
        for member_id in row["member_ids"]:
            member_to_row[member_id] = row
    return member_to_row, canonical_by_id


def _official_sales_stage_max_date():
    return get_official_sales_stage_max_date()


def _recent_sales_stage_max_date():
    return get_recent_sales_stage_max_date()


def _authoritative_sales_max_date():
    return canonical_authoritative_sales_max_date()


def _v2_category_sales_max_date():
    return canonical_v2_category_sales_max_date()


def _v2_product_sales_max_date():
    return canonical_v2_product_sales_max_date()


def _canonical_point_max_date():
    return get_canonical_point_max_date()


def _canonical_point_previous_dates(target_date) -> list:
    return get_canonical_point_previous_dates(target_date)


def _operational_sales_filters(*, start_date, end_date) -> Q:
    return canonical_operational_sales_filters(start_date=start_date, end_date=end_date)


def _operational_sales_rows_for_date(target_date):
    return canonical_operational_sales_rows_for_date(target_date)


def _sales_source_context() -> dict[str, object]:
    return get_canonical_sales_source_context(
        cache_key_parts=("erp", "bi", "sales-source-context"),
    )


def _build_sales_source_context() -> dict[str, object]:
    return build_canonical_sales_source_context()


def _official_daily_cut_for_date(target_date: date | None):
    if not target_date:
        return None
    return CorteOficialDiario.objects.filter(corte_date=target_date).first()


def _branch_control_action_url(*, sucursal_id: int | None, target_date: date | None) -> str:
    if sucursal_id and target_date:
        return reverse("control:discrepancias") + f"?{urlencode({'from': target_date.isoformat(), 'to': target_date.isoformat(), 'sucursal_id': sucursal_id})}"
    return reverse("reportes:ventas")


def _financial_product_action_url(recipe_name: str | None, *, bucket: str | None = None) -> str:
    query: dict[str, object] = {}
    if recipe_name:
        query["q"] = recipe_name
    if bucket:
        query["bucket"] = bucket
    return reverse("reportes:financiero") + (f"?{urlencode(query)}" if query else "")


def _sales_rows_for_date(source: dict[str, object], target_date):
    return canonical_sales_rows_for_date(source, target_date)


def _sales_rows_for_month(source: dict[str, object], year: int, month: int):
    return canonical_sales_rows_for_month(source, year, month)


def _point_sales_month_total(year: int, month: int) -> dict[str, object]:
    return canonical_point_sales_month_total(year, month)


def _sales_previous_dates(source: dict[str, object], target_date) -> list:
    return canonical_sales_previous_dates(source, target_date)


def _sales_history_queryset(source: dict[str, object]):
    return canonical_sales_history_queryset(source)


def _format_periodic_schedule(task) -> str:
    if task is None:
        return "Sin scheduler configurado"
    if getattr(task, "crontab_id", None) and task.crontab:
        cron = task.crontab
        return f"Diario {cron.hour.zfill(2)}:{cron.minute.zfill(2)} ({cron.timezone})"
    if getattr(task, "interval_id", None) and task.interval:
        return f"Cada {task.interval.every} {task.interval.period}"
    return "Programación no identificada"


def _coerce_audit_reference_date(audit) -> date | None:
    if not audit:
        return None
    payload = dict(getattr(audit, "payload", {}) or {})
    raw_value = payload.get("reference_date") or payload.get("date")
    if not raw_value:
        return None
    try:
        return date.fromisoformat(str(raw_value))
    except ValueError:
        return None


def _expected_sales_cut_date(task) -> dict[str, object]:
    scheduler_tz = timezone.get_current_timezone()
    scheduler_tz_label = str(getattr(scheduler_tz, "key", "") or scheduler_tz)
    schedule_hour = 3
    schedule_minute = 35

    if task is not None and getattr(task, "crontab_id", None) and task.crontab:
        cron = task.crontab
        cron_hour = str(cron.hour or "").strip()
        cron_minute = str(cron.minute or "").strip()
        if cron_hour.isdigit():
            schedule_hour = int(cron_hour)
        if cron_minute.isdigit():
            schedule_minute = int(cron_minute)
        scheduler_tz_label = str(cron.timezone or scheduler_tz_label)
        try:
            scheduler_tz = ZoneInfo(scheduler_tz_label)
        except ZoneInfoNotFoundError:
            scheduler_tz = timezone.get_current_timezone()

    scheduler_now = timezone.localtime(timezone=scheduler_tz)
    cutoff_has_run = (scheduler_now.hour, scheduler_now.minute) >= (schedule_hour, schedule_minute)
    expected_cut_date = scheduler_now.date() - timedelta(days=1 if cutoff_has_run else 2)
    return {
        "expected_cut_date": expected_cut_date,
        "schedule_hour": schedule_hour,
        "schedule_minute": schedule_minute,
        "timezone_label": scheduler_tz_label,
        "schedule_time_label": f"{schedule_hour:02d}:{schedule_minute:02d}",
    }


def _sales_refresh_status(*, visible_cut_date: date | None = None) -> dict[str, object]:
    scheduler_task = None
    try:
        from django_celery_beat.models import PeriodicTask

        scheduler_task = PeriodicTask.objects.filter(name="reportes: refresh analytics operativo").select_related(
            "crontab",
            "interval",
        ).first()
    except Exception:
        scheduler_task = None

    latest_completed = (
        AuditLog.objects.filter(action="INTEGRATIONS_OPERATIONAL_REFRESH_COMPLETED", model="reportes.AnalyticRefreshWindow")
        .select_related("user")
        .order_by("-timestamp", "-id")
        .first()
    )
    latest_failed = (
        AuditLog.objects.filter(action="INTEGRATIONS_OPERATIONAL_REFRESH_FAILED", model="reportes.AnalyticRefreshWindow")
        .select_related("user")
        .order_by("-timestamp", "-id")
        .first()
    )
    latest_requested = (
        AuditLog.objects.filter(action="REPORTES_BI_FORCE_REFRESH_REQUESTED", model="reportes.AnalyticRefreshWindow")
        .select_related("user")
        .order_by("-timestamp", "-id")
        .first()
    )
    expected_meta = _expected_sales_cut_date(scheduler_task)
    expected_cut_date = expected_meta["expected_cut_date"]
    expected_cut_date_iso = expected_cut_date.isoformat() if expected_cut_date else ""
    visible_cut_date_iso = visible_cut_date.isoformat() if visible_cut_date else ""

    latest_terminal = latest_completed
    if latest_failed and (
        latest_terminal is None or (latest_failed.timestamp, latest_failed.id) >= (latest_terminal.timestamp, latest_terminal.id)
    ):
        latest_terminal = latest_failed

    lock_active = cache.get(BI_FORCE_REFRESH_LOCK_KEY) is not None
    requested_after_terminal = bool(
        latest_requested
        and (
            latest_terminal is None
            or (latest_requested.timestamp, latest_requested.id) >= (latest_terminal.timestamp, latest_terminal.id)
        )
    )
    request_target_date = _coerce_audit_reference_date(latest_requested)
    completed_target_date = _coerce_audit_reference_date(latest_completed)
    failed_target_date = _coerce_audit_reference_date(latest_failed)
    request_matches_visible = bool(
        request_target_date
        and visible_cut_date
        and request_target_date == visible_cut_date
    )
    pending_recent = bool(
        requested_after_terminal
        and latest_requested
        and (request_matches_visible or visible_cut_date is None)
        and (
            lock_active
            or (timezone.now() - latest_requested.timestamp).total_seconds() <= BI_FORCE_REFRESH_LOCK_SECONDS
        )
    )
    stale_pending = bool(requested_after_terminal and latest_requested and not pending_recent)
    is_cut_delayed = bool(visible_cut_date and expected_cut_date and visible_cut_date < expected_cut_date)
    cut_lag_days = (expected_cut_date - visible_cut_date).days if is_cut_delayed else 0
    target_refresh_date = expected_cut_date or visible_cut_date or timezone.localdate()

    if latest_terminal is latest_failed and not requested_after_terminal:
        last_status = "ERROR"
        last_event = latest_failed
        status_title = "La actualización falló"
        if is_cut_delayed:
            status_message = (
                f"El corte visible sigue en {visible_cut_date_iso} y debería estar al menos en {expected_cut_date_iso}."
            )
        else:
            status_message = "Revisa la bitácora o vuelve a intentarlo cuando el proceso operativo esté disponible."
        action_label = "Reintentar actualización"
        status_label = "Error"
    elif is_cut_delayed and pending_recent:
        last_status = "PENDIENTE_REZAGO"
        last_event = latest_requested
        status_title = "Pendiente con rezago"
        status_message = (
            f"Fecha visible {visible_cut_date_iso} · fecha esperada {expected_cut_date_iso}. "
            f"El refresh solicitado busca cerrar ese rezago."
        )
        action_label = "Actualización en curso"
        status_label = "Pendiente con rezago"
    elif is_cut_delayed and latest_completed and completed_target_date and completed_target_date >= expected_cut_date:
        last_status = "INCONSISTENCIA"
        last_event = latest_completed
        status_title = "Inconsistencia detectada"
        status_message = (
            f"La automatización terminó para {completed_target_date.isoformat()}, pero el corte visible sigue en {visible_cut_date_iso}."
        )
        action_label = f"Actualizar hasta {expected_cut_date_iso}"
        status_label = "Inconsistencia detectada"
    elif is_cut_delayed and stale_pending:
        last_status = "INCONSISTENCIA"
        last_event = latest_requested
        status_title = "Pendiente estancado"
        status_message = (
            f"Se solicitó actualización para {request_target_date.isoformat() if request_target_date else expected_cut_date_iso}, "
            f"pero el corte visible sigue atrasado en {cut_lag_days} día(s)."
        )
        action_label = f"Actualizar hasta {expected_cut_date_iso}"
        status_label = "Inconsistencia detectada"
    elif is_cut_delayed:
        last_status = "REZAGO"
        last_event = latest_requested or latest_completed or latest_failed
        status_title = "Rezago detectado"
        status_message = (
            f"Fecha visible {visible_cut_date_iso} · fecha esperada {expected_cut_date_iso}. "
            f"Falta cerrar {cut_lag_days} día(s)."
        )
        action_label = f"Actualizar hasta {expected_cut_date_iso}"
        status_label = "Rezago detectado"
    elif pending_recent:
        last_status = "PENDIENTE"
        last_event = latest_requested
        status_title = "Pendiente sin rezago"
        status_message = (
            f"Se está revalidando el corte visible {request_target_date.isoformat() if request_target_date else (visible_cut_date_iso or expected_cut_date_iso)}. "
            "No hace falta volver a solicitar la actualización."
        )
        action_label = "Actualización en curso"
        status_label = "Pendiente sin rezago"
    elif latest_completed:
        last_status = "OK"
        last_event = latest_completed
        status_title = "Al día"
        status_message = (
            f"Fecha visible {visible_cut_date_iso or expected_cut_date_iso} · fecha esperada {expected_cut_date_iso}. "
            "No hay rezago operativo."
        )
        action_label = "Actualizar ventas"
        status_label = "Al día"
    else:
        last_status = "SIN_EJECUCION"
        last_event = None
        status_title = "Sin ejecuciones auditadas"
        if is_cut_delayed:
            status_message = (
                f"Fecha visible {visible_cut_date_iso} · fecha esperada {expected_cut_date_iso}. "
                "Aún no hay una ejecución auditada que cierre ese rezago."
            )
        else:
            status_message = "El botón encola la actualización del corte visible y el worker la procesa en segundo plano."
        action_label = "Actualizar ventas"
        status_label = "Sin ejecución"

    return {
        "configured": scheduler_task is not None,
        "enabled": bool(getattr(scheduler_task, "enabled", False)),
        "schedule_label": _format_periodic_schedule(scheduler_task),
        "scheduler_name": getattr(scheduler_task, "name", ""),
        "lock_active": lock_active,
        "is_refresh_pending": pending_recent,
        "button_disabled": lock_active,
        "status_title": status_title,
        "status_message": status_message,
        "status_label": status_label,
        "button_label": action_label,
        "last_status": last_status,
        "last_event": last_event,
        "last_completed": latest_completed,
        "last_failed": latest_failed,
        "last_requested": latest_requested,
        "last_refresh_timestamp": getattr(last_event, "timestamp", None),
        "visible_cut_date": visible_cut_date,
        "visible_cut_date_iso": visible_cut_date_iso,
        "expected_cut_date": expected_cut_date,
        "expected_cut_date_iso": expected_cut_date_iso,
        "target_refresh_date": target_refresh_date,
        "target_refresh_date_iso": target_refresh_date.isoformat() if target_refresh_date else "",
        "cut_lag_days": cut_lag_days,
        "is_cut_delayed": is_cut_delayed,
        "stale_pending": stale_pending,
        "request_target_date": request_target_date.isoformat() if request_target_date else "",
        "completed_target_date": completed_target_date.isoformat() if completed_target_date else "",
        "failed_target_date": failed_target_date.isoformat() if failed_target_date else "",
        "schedule_time_label": expected_meta["schedule_time_label"],
        "schedule_timezone_label": expected_meta["timezone_label"],
    }


def _canonical_sales_history_summary(source: dict[str, object]) -> dict[str, object] | None:
    canonical_latest = source.get("canonical_latest_date")
    if not canonical_latest:
        return None

    authoritative_first = VentaAutoritativaPoint.objects.order_by("sale_date").values_list("sale_date", flat=True).first()
    v2_category_first = PointSalesDailyCategoryFact.objects.order_by("sale_date").values_list("sale_date", flat=True).first()
    v2_product_first = PointSalesDailyProductFact.objects.order_by("sale_date").values_list("sale_date", flat=True).first()
    start_candidates = [value for value in [authoritative_first, v2_category_first, v2_product_first] if value]
    if not start_candidates:
        return None

    start_date = min(start_candidates)
    selected = get_sales_range(
        start_date=start_date,
        end_date=canonical_latest,
        coverage_policy="prefer_complete",
    )
    if selected["source"] == "none":
        return None

    if selected["source"] == "authoritative":
        rows_qs = VentaAutoritativaPoint.objects.all()
        total_rows = rows_qs.count()
        first_date = rows_qs.order_by("sale_date").values_list("sale_date", flat=True).first()
        last_date = rows_qs.order_by("-sale_date").values_list("sale_date", flat=True).first()
        recipe_count = rows_qs.exclude(product_id__isnull=True).values_list("product_id", flat=True).distinct().count()
        top_branches = list(rows_qs.values("branch__codigo", "branch__nombre").annotate(total=Sum("quantity")).order_by("-total", "branch__codigo")[:4])
        top_recipes = list(rows_qs.values("product__nombre", "point_name").annotate(total=Sum("quantity")).order_by("-total", "product__nombre", "point_name")[:5])
    else:
        category_qs = PointSalesDailyCategoryFact.objects.all()
        product_qs = PointSalesDailyProductFact.objects.all()
        total_rows = category_qs.count()
        first_date = category_qs.order_by("sale_date").values_list("sale_date", flat=True).first()
        last_date = category_qs.order_by("-sale_date").values_list("sale_date", flat=True).first()
        recipe_count = product_qs.exclude(receta_id__isnull=True).values_list("receta_id", flat=True).distinct().count() if product_qs.exists() else 0
        top_branches = list(category_qs.values("branch__erp_branch__codigo", "branch__erp_branch__nombre").annotate(total=Sum("total_cantidad")).order_by("-total", "branch__erp_branch__codigo")[:4])
        top_recipes = list(product_qs.values("receta__nombre", "producto_nombre_historico").annotate(total=Sum("total_cantidad")).order_by("-total", "receta__nombre", "producto_nombre_historico")[:5]) if product_qs.exists() else []

    active_days = int(selected.get("coverage_days") or 0)
    branch_count = int(selected.get("coverage_branches") or 0)
    expected_days = ((last_date - first_date).days + 1) if first_date and last_date else 0
    missing_days = max(expected_days - active_days, 0)
    return {
        "available": True,
        "status": "Cobertura cerrada" if missing_days == 0 else "Cobertura parcial",
        "tone": "success" if missing_days == 0 else "warning",
        "official_ready": missing_days == 0,
        "detail": (
            "Fuente canónica Point disponible para lectura ejecutiva."
            if missing_days == 0
            else f"Fuente canónica Point disponible con {missing_days} día(s) faltantes dentro del rango visible."
        ),
        "source_label": "Point directo",
        "date_label": f"{first_date.strftime('%d/%m/%Y')} → {last_date.strftime('%d/%m/%Y')}" if first_date and last_date else "Sin cobertura",
        "active_days": active_days,
        "expected_days": expected_days,
        "missing_days": missing_days,
        "branch_count": branch_count,
        "recipe_count": recipe_count,
        "total_rows": total_rows,
        "total_units": Decimal(str(selected.get("cantidad") or 0)),
        "total_amount": Decimal(str(selected.get("monto") or 0)),
        "top_branches": top_branches,
        "top_recipes": top_recipes,
    }


def _ventas_historicas_bi_summary() -> dict[str, object]:
    source = _sales_source_context()
    rows_qs = _sales_history_queryset(source)
    total_rows = rows_qs.count()
    if total_rows == 0:
        if source["mode"] == "point_stage" and source.get("canonical_latest_date"):
            canonical_summary = _canonical_sales_history_summary(source)
            if canonical_summary is not None:
                return canonical_summary
        return {
            "available": False,
            "status": "Sin histórico",
            "tone": "warning",
            "detail": "No hay ventas históricas cargadas para análisis ejecutivo.",
            "source_label": source["label"],
            "date_label": "Sin cobertura",
            "active_days": 0,
            "expected_days": 0,
            "missing_days": 0,
            "branch_count": 0,
            "recipe_count": 0,
            "total_rows": 0,
            "total_units": Decimal("0"),
            "total_amount": Decimal("0"),
            "top_branches": [],
            "top_recipes": [],
        }
    if source["mode"] == "point_stage":
        total_units = rows_qs.aggregate(total=Sum("quantity")).get("total") or Decimal("0")
        total_amount = rows_qs.aggregate(total=Sum("total_amount")).get("total") or Decimal("0")
        first_date = rows_qs.order_by("sale_date").values_list("sale_date", flat=True).first()
        last_date = rows_qs.order_by("-sale_date").values_list("sale_date", flat=True).first()
        active_days = rows_qs.values_list("sale_date", flat=True).distinct().count()
        branch_count = rows_qs.values_list("branch_id", flat=True).distinct().count()
        recipe_count = rows_qs.values_list("product_id", flat=True).distinct().count()
        top_branches = list(rows_qs.values("branch__external_id", "branch__name").annotate(total=Sum("quantity")).order_by("-total", "branch__external_id")[:4])
        top_recipes = list(rows_qs.values("product__name").annotate(total=Sum("quantity")).order_by("-total", "product__name")[:5])
    else:
        total_units = rows_qs.aggregate(total=Sum("cantidad")).get("total") or Decimal("0")
        total_amount = rows_qs.aggregate(total=Sum("monto_total")).get("total") or Decimal("0")
        first_date = rows_qs.order_by("fecha").values_list("fecha", flat=True).first()
        last_date = rows_qs.order_by("-fecha").values_list("fecha", flat=True).first()
        active_days = rows_qs.values_list("fecha", flat=True).distinct().count()
        branch_count = rows_qs.exclude(sucursal_id__isnull=True).values_list("sucursal_id", flat=True).distinct().count()
        recipe_count = rows_qs.values_list("receta_id", flat=True).distinct().count()
        top_branches = list(rows_qs.exclude(sucursal_id__isnull=True).values("sucursal__codigo", "sucursal__nombre").annotate(total=Sum("cantidad")).order_by("-total", "sucursal__codigo")[:4])
        top_recipes = list(rows_qs.values("receta__nombre").annotate(total=Sum("cantidad")).order_by("-total", "receta__nombre")[:5])
    expected_days = ((last_date - first_date).days + 1) if first_date and last_date else 0
    missing_days = max(expected_days - active_days, 0)
    detail = (
        f"{source['detail']} La serie diaria está lista para lectura ejecutiva y planeación."
        if missing_days == 0
        else f"{source['detail']} Hay {missing_days} día(s) faltantes dentro del rango histórico cargado."
    )
    if (
        source["mode"] == "point_stage"
        and source.get("canonical_latest_date")
        and source.get("stage_latest_date")
        and source["canonical_latest_date"] > source["stage_latest_date"]
    ):
        detail += f" La última fecha canónica disponible es {source['canonical_latest_date'].strftime('%d/%m/%Y')}, pero la serie visible legacy llega a {source['stage_latest_date'].strftime('%d/%m/%Y')}."
    return {
        "available": True,
        "status": "Cobertura cerrada" if missing_days == 0 else "Cobertura parcial",
        "tone": "success" if missing_days == 0 and source["canonical"] else "warning",
        "official_ready": bool(source["canonical"] and missing_days == 0),
        "detail": detail,
        "source_label": source["label"],
        "date_label": f"{first_date.strftime('%d/%m/%Y')} → {last_date.strftime('%d/%m/%Y')}" if first_date and last_date else "Sin cobertura",
        "active_days": active_days,
        "expected_days": expected_days,
        "missing_days": missing_days,
        "branch_count": branch_count,
        "recipe_count": recipe_count,
        "total_rows": total_rows,
        "total_units": total_units,
        "total_amount": total_amount,
        "top_branches": top_branches,
        "top_recipes": top_recipes,
    }


def _bi_daily_sales_snapshot() -> dict[str, object]:
    dataset = get_dashboard_sales_dataset()
    snapshot = dict(dataset.get("daily_sales_snapshot") or {})
    latest_date = snapshot.get("date")
    if not latest_date:
        return {
            "status": "Sin cortes",
            "tone": "warning",
            "detail": "Todavía no hay ventas recientes para lectura ejecutiva diaria.",
            "date_label": "Sin fecha",
            "source_label": "Sin fuente",
            "total_units": Decimal("0"),
            "total_amount": Decimal("0"),
            "total_tickets": 0,
            "branch_count": 0,
            "recipe_count": 0,
            "comparison_label": "Sin comparativo",
            "comparison_tone": "warning",
            "comparison_detail": "Carga cortes de venta para habilitar lectura diaria.",
            "comparison_basis": "Sin referencia disponible",
            "top_branches": [],
            "top_products": [],
            "required_branch_count": 0,
            "present_required_branch_count": 0,
            "missing_required_branch_count": 0,
            "missing_required_branches": [],
            "completeness_label": "Sin validación",
        }

    required_branches = _required_daily_sales_branches(latest_date)
    present_branch_ids = set(dataset.get("present_branch_ids") or [])
    present_branch_ids.update(_indicator_present_daily_sales_branch_ids(latest_date))
    present_branch_ids.update(_official_backfill_present_daily_sales_branch_ids(latest_date))
    validated_zero_codes = _validated_zero_daily_sales_branch_codes(latest_date)
    missing_required_branches = [
        branch
        for branch in required_branches
        if int(branch["branch_id"]) not in present_branch_ids
        and str(branch.get("branch_code") or "") not in validated_zero_codes
    ]
    missing_required_branch_count = len(missing_required_branches)
    required_branch_count = len(required_branches)
    present_required_branch_count = max(required_branch_count - missing_required_branch_count, 0)
    completeness_label = (
        "Corte completo"
        if missing_required_branch_count == 0
        else f"Faltan {missing_required_branch_count} sucursal(es) obligatoria(s)"
    )
    missing_branch_names = [str(branch.get("branch_name") or branch.get("branch_code") or "Sucursal") for branch in missing_required_branches]
    official_cut = _official_daily_cut_for_date(latest_date)

    source_label = "Point directo"
    status = "Corte cargado"
    tone = "success"
    detail = "Resumen del último corte disponible. Fuente canónica Point bridge."
    if official_cut:
        source_label = official_cut.source_label or "Corte oficial diario"
        status = "Corte oficial conciliado"
        detail = "El total visible usa un corte oficial conciliado contra la operación del día."
    if missing_required_branch_count and not official_cut:
        preview = ", ".join(missing_branch_names[:4])
        if missing_required_branch_count > 4:
            preview = f"{preview} y {missing_required_branch_count - 4} más"
        status = "Corte incompleto"
        tone = "danger"
        detail = (
            f"Faltan sucursales obligatorias en el corte diario: {preview}. "
            "El total visible no debe considerarse cierre válido hasta completar esas sucursales."
        )
    elif missing_required_branch_count and official_cut:
        preview = ", ".join(missing_branch_names[:4])
        if missing_required_branch_count > 4:
            preview = f"{preview} y {missing_required_branch_count - 4} más"
        tone = "warning"
        detail = f"El total visible usa corte oficial conciliado. Point crudo sigue incompleto y no trae: {preview}."

    snapshot.update(
        {
            "status": status,
            "tone": tone,
            "detail": detail,
            "source_label": source_label,
            "required_branch_count": required_branch_count,
            "present_required_branch_count": present_required_branch_count,
            "missing_required_branch_count": missing_required_branch_count,
            "missing_required_branches": missing_required_branches,
            "missing_required_branch_names": missing_branch_names,
            "completeness_label": completeness_label,
            "official_cut_applied": bool(official_cut),
            "official_cut_evidence_path": official_cut.evidence_path if official_cut else "",
            "official_cut_notes": official_cut.notes if official_cut else "",
            "indicator_total_amount": snapshot.get("raw_total_amount") or Decimal("0"),
            "top_branches": [
                {
                    **row,
                    "action_url": _branch_control_action_url(
                        sucursal_id=row.get("branch_id"),
                        target_date=latest_date,
                    ),
                }
                for row in list(snapshot.get("top_branches") or [])
            ],
            "top_products": [
                {
                    **row,
                    "action_url": _financial_product_action_url(row.get("recipe_name") or row.get("label")),
                }
                for row in list(snapshot.get("top_products") or [])
            ],
            "mapping_coverage_pct": Decimal("100") if snapshot.get("total_amount") else None,
            "mapped_amount": snapshot.get("total_amount") or Decimal("0"),
            "mapped_units": snapshot.get("total_units") or Decimal("0"),
            "unmapped_amount": Decimal("0"),
            "unmapped_units": Decimal("0"),
        }
    )
    return snapshot


def _bi_branch_weekday_comparisons(limit: int = 5) -> list[dict[str, object]]:
    source = _sales_source_context()
    latest_date = source["latest_date"]
    if not latest_date:
        return []
    comparable_date = next((date_value for date_value in _sales_previous_dates(source, latest_date) if date_value.weekday() == latest_date.weekday()), None)
    if not comparable_date:
        return []
    if source["mode"] == "point_stage":
        bulk = get_daily_sales_bulk(
            fechas=[latest_date, comparable_date],
            dimension="branch",
            include_indicators=True,
            coverage_policy="strict_priority",
        )
        current_payload = bulk["dates"].get(latest_date.isoformat(), {})
        comparable_payload = bulk["dates"].get(comparable_date.isoformat(), {})
        current_rows = list(current_payload.get("rows") or [])
        current_indicator_map = current_payload.get("indicator_map") or {}
        comparable_map = {
            row["branch_id"]: row
            for row in comparable_payload.get("rows") or []
        }
        comparable_indicator_map = comparable_payload.get("indicator_map") or {}
    else:
        current_rows = list(_sales_rows_for_date(source, latest_date).exclude(sucursal_id__isnull=True).values("sucursal_id", "sucursal__codigo", "sucursal__nombre").annotate(units=Sum("cantidad"), amount=Sum("monto_total"), tickets=Sum("tickets")))
        comparable_map = {row["sucursal_id"]: row for row in _sales_rows_for_date(source, comparable_date).exclude(sucursal_id__isnull=True).values("sucursal_id").annotate(units=Sum("cantidad"), amount=Sum("monto_total"), tickets=Sum("tickets"))}
    rows: list[dict[str, object]] = []
    for row in current_rows:
        branch_id = row["branch_id"] if source["mode"] == "point_stage" else row["sucursal_id"]
        comparable = comparable_map.get(branch_id)
        if not comparable:
            continue
        if source["mode"] == "point_stage":
            current_indicator = current_indicator_map.get(branch_id) or {}
            comparable_indicator = comparable_indicator_map.get(branch_id) or {}
            current_amount = Decimal(str(current_indicator.get("amount") or row.get("amount") or 0))
            comparable_amount = Decimal(str(comparable_indicator.get("amount") or comparable.get("amount") or 0))
            current_tickets = int(current_indicator.get("tickets") or row.get("tickets") or 0)
        else:
            current_amount = Decimal(str(row.get("amount") or 0))
            comparable_amount = Decimal(str(comparable.get("amount") or 0))
            current_tickets = int(row.get("tickets") or 0)
        current_units = Decimal(str(row.get("units") or 0))
        comparable_units = Decimal(str(comparable.get("units") or 0))
        delta_pct = ((current_amount - comparable_amount) / comparable_amount) * Decimal("100") if comparable_amount > 0 else (((current_units - comparable_units) / comparable_units) * Decimal("100") if comparable_units > 0 else None)
        if delta_pct is None:
            continue
        if delta_pct <= Decimal("-12"):
            status, tone, detail, rank_score = "Abajo del comparable", "danger", f"-{abs(delta_pct):.1f}% vs {comparable_date.isoformat()}", abs(delta_pct) + Decimal("100")
        elif delta_pct >= Decimal("12"):
            status, tone, detail, rank_score = "Arriba del comparable", "success", f"+{delta_pct:.1f}% vs {comparable_date.isoformat()}", delta_pct
        else:
            status, tone, detail, rank_score = "Dentro de rango", "warning", f"{delta_pct:.1f}% vs {comparable_date.isoformat()}", abs(delta_pct)
        rows.append({
            "branch_code": row.get("branch_code") or row.get("branch__external_id") or row.get("sucursal__codigo") or "SIN-COD",
            "branch_name": row.get("branch_name") or row.get("branch__name") or row.get("sucursal__nombre") or "Sucursal",
            "units": current_units,
            "amount": current_amount,
            "tickets": current_tickets,
            "status": status,
            "tone": tone,
            "detail": detail,
            "rank_score": rank_score,
            "delta_pct": delta_pct,
            "action_url": _branch_control_action_url(
                sucursal_id=row.get("erp_branch_id") or row.get("branch__erp_branch_id") or row.get("sucursal_id"),
                target_date=latest_date,
            ),
        })
    severity_order = {"danger": 0, "warning": 1, "success": 2}
    rows.sort(key=lambda item: (severity_order.get(str(item.get("tone") or ""), 9), -float(item.get("rank_score") or 0)))
    return rows[:limit]


def _bi_product_weekday_comparisons(limit: int = 5) -> list[dict[str, object]]:
    source = _sales_source_context()
    latest_date = source["latest_date"]
    if not latest_date:
        return []
    comparable_date = next((date_value for date_value in _sales_previous_dates(source, latest_date) if date_value.weekday() == latest_date.weekday()), None)
    if not comparable_date:
        return []
    if source["mode"] == "point_stage":
        bulk = get_daily_sales_bulk(
            fechas=[latest_date, comparable_date],
            dimension="product",
            coverage_policy="strict_priority",
        )
        current_payload = bulk["dates"].get(latest_date.isoformat(), {})
        comparable_payload = bulk["dates"].get(comparable_date.isoformat(), {})
        current_rows = list(current_payload.get("rows") or [])
        comparable_map = {
            row["key"]: row
            for row in comparable_payload.get("rows") or []
        }
    else:
        current_rows = list(_sales_rows_for_date(source, latest_date).exclude(receta_id__isnull=True).filter(monto_total__gt=0).values("receta_id", "receta__nombre").annotate(units=Sum("cantidad"), amount=Sum("monto_total"), branch_count=Count("sucursal", distinct=True)))
        comparable_map = {row["receta_id"]: row for row in _sales_rows_for_date(source, comparable_date).exclude(receta_id__isnull=True).filter(monto_total__gt=0).values("receta_id").annotate(units=Sum("cantidad"), amount=Sum("monto_total"), branch_count=Count("sucursal", distinct=True))}
    rows: list[dict[str, object]] = []
    for row in current_rows:
        product_id = row["key"] if source["mode"] == "point_stage" else row["receta_id"]
        comparable = comparable_map.get(product_id)
        if not comparable:
            continue
        current_amount = Decimal(str(row.get("amount") or 0))
        current_units = Decimal(str(row.get("units") or 0))
        comparable_amount = Decimal(str(comparable.get("amount") or 0))
        comparable_units = Decimal(str(comparable.get("units") or 0))
        delta_pct = ((current_amount - comparable_amount) / comparable_amount) * Decimal("100") if comparable_amount > 0 else (((current_units - comparable_units) / comparable_units) * Decimal("100") if comparable_units > 0 else None)
        if delta_pct is None:
            continue
        if delta_pct <= Decimal("-15"):
            status, tone, detail, rank_score = "Abajo del comparable", "danger", f"-{abs(delta_pct):.1f}% vs {comparable_date.isoformat()}", abs(delta_pct) + Decimal("100")
        elif delta_pct >= Decimal("15"):
            status, tone, detail, rank_score = "Arriba del comparable", "success", f"+{delta_pct:.1f}% vs {comparable_date.isoformat()}", delta_pct
        else:
            status, tone, detail, rank_score = "Dentro de rango", "warning", f"{delta_pct:.1f}% vs {comparable_date.isoformat()}", abs(delta_pct)
        product_label = row.get("recipe_name") or row.get("product_name") or row.get("receta__nombre") or row.get("product__name") or "Producto"
        rows.append({
            "recipe_name": product_label,
            "units": current_units,
            "amount": current_amount,
            "branch_count": int(row.get("branch_count") or 0),
            "status": status,
            "tone": tone,
            "detail": detail,
            "rank_score": rank_score,
            "delta_pct": delta_pct,
            "action_url": _financial_product_action_url(product_label),
        })
    severity_order = {"danger": 0, "warning": 1, "success": 2}
    rows.sort(key=lambda item: (severity_order.get(str(item.get("tone") or ""), 9), -float(item.get("rank_score") or 0)))
    return rows[:limit]


@login_required
def ventas(request: HttpRequest) -> HttpResponse:
    if not can_view_reportes(request.user):
        raise PermissionDenied("No tienes permisos para ver Reportes.")

    try:
        months_window = int(request.GET.get("months") or "6")
    except (TypeError, ValueError):
        months_window = 6

    daily_sales_snapshot = _bi_daily_sales_snapshot()
    forecast_panel = build_sales_forecast_panel()
    yoy_panel = build_monthly_yoy_panel(months=months_window)
    branch_weekday_rows = _bi_branch_weekday_comparisons(limit=6)
    product_weekday_rows = _bi_product_weekday_comparisons(limit=6)
    ventas_historicas_summary = _ventas_historicas_bi_summary()
    visible_cut_date = daily_sales_snapshot.get("date")
    refresh_status = _sales_refresh_status(visible_cut_date=visible_cut_date)
    refresh_reference_date = refresh_status.get("target_refresh_date_iso") or (
        visible_cut_date.isoformat() if visible_cut_date else timezone.localdate().isoformat()
    )

    context = {
        "daily_sales_snapshot": daily_sales_snapshot,
        "forecast_panel": forecast_panel,
        "yoy_panel": yoy_panel,
        "branch_weekday_rows": branch_weekday_rows,
        "product_weekday_rows": product_weekday_rows,
        "ventas_historicas_summary": ventas_historicas_summary,
        "months_window": months_window,
        "can_manage_sales_refresh": can_manage_orquestacion(request.user),
        "sales_refresh_status": refresh_status,
        "sales_refresh_reference_date": refresh_reference_date,
        "sales_refresh_lookback_days": 7,
        "sales_branch_bar_rows": [
            {
                **row,
                "label": row.get("secondary") or row.get("label") or "Sucursal",
                "secondary": row.get("label") if row.get("secondary") else "",
            }
            for row in _bi_bar_rows(
                list(daily_sales_snapshot.get("top_branches") or []),
                label_key="label",
                secondary_key="secondary",
                value_key="amount",
            )
        ],
        "sales_product_bar_rows": _bi_bar_rows(
            list(daily_sales_snapshot.get("top_products") or []),
            label_key="label",
            value_key="amount",
        ),
        "branch_weekday_bar_rows": [
            {
                **row,
                "label": row.get("secondary") or row.get("label") or "Sucursal",
                "secondary": row.get("label") if row.get("secondary") else "",
            }
            for row in _bi_comparison_bar_rows(
                list(branch_weekday_rows or []),
                label_key="branch_code",
                secondary_key="branch_name",
            )
        ],
        "product_weekday_bar_rows": _bi_comparison_bar_rows(
            list(product_weekday_rows or []),
            label_key="recipe_name",
        ),
    }
    return render(request, "reportes/ventas.html", context)


@login_required
def cierre_operativo(request: HttpRequest) -> HttpResponse:
    if not can_view_reportes(request.user):
        raise PermissionDenied("No tienes permisos para ver Reportes.")

    raw_target_date = (request.GET.get("fecha") or "").strip()
    expected_meta = _expected_sales_cut_date(None)
    default_target_date = expected_meta.get("expected_cut_date") or (timezone.localdate() - timedelta(days=1))
    target_date = default_target_date
    invalid_date = False
    if raw_target_date:
        try:
            target_date = date.fromisoformat(raw_target_date)
        except ValueError:
            invalid_date = True

    closure = build_daily_operational_closure(target_date=target_date)
    daily_sales_snapshot = _bi_daily_sales_snapshot()
    production_snapshot = _bi_production_snapshot()
    waste_summary = _bi_waste_summary(target_date, target_date)
    inventory_snapshot = _bi_inventory_snapshot()

    closure["sales"]["visible_cut_snapshot_date"] = daily_sales_snapshot.get("date")
    closure["sales"]["visible_cut_snapshot_label"] = daily_sales_snapshot.get("date_label") or ""
    closure["production"]["plan_snapshot"] = production_snapshot
    closure["waste"]["visible_summary"] = waste_summary
    closure["inventory"]["visible_snapshot"] = inventory_snapshot

    date_options: list[dict[str, str]] = []
    seen_dates: set[date] = set()
    candidate_dates = [
        target_date,
        default_target_date,
        daily_sales_snapshot.get("date"),
        closure["sales"].get("raw_max_date"),
        closure["production"].get("raw_max_date"),
        closure["waste"].get("raw_max_date"),
        closure["inventory"].get("fact_max_date"),
    ]
    for candidate in candidate_dates:
        if not candidate or candidate in seen_dates:
            continue
        seen_dates.add(candidate)
        date_options.append({"value": candidate.isoformat(), "label": candidate.strftime("%Y-%m-%d")})
    date_options.sort(key=lambda item: item["value"], reverse=True)

    if invalid_date:
        messages.warning(request, "La fecha solicitada no es válida. Se mostró la fecha operativa esperada.")

    context = {
        "closure": closure,
        "target_date": target_date,
        "target_date_iso": target_date.isoformat(),
        "default_target_date_iso": default_target_date.isoformat(),
        "date_options": date_options,
        "module_tabs": _reportes_module_tabs("cierre_operativo"),
        "daily_sales_snapshot": daily_sales_snapshot,
        "production_snapshot": production_snapshot,
        "waste_summary": waste_summary,
        "inventory_snapshot": inventory_snapshot,
    }
    return render(request, "reportes/cierre_operativo_diario.html", context)


def _bi_bar_rows(
    raw_rows: list[dict[str, object]],
    label_key: str,
    value_key: str,
    secondary_key: str | None = None,
    limit: int = 6,
) -> list[dict[str, object]]:
    rows = list(raw_rows or [])[:limit]
    max_value = max((Decimal(str(item.get(value_key) or 0)) for item in rows), default=Decimal("0"))
    output: list[dict[str, object]] = []
    for item in rows:
        value = Decimal(str(item.get(value_key) or 0))
        pct = float((value / max_value) * Decimal("100")) if max_value > 0 else 0.0
        output.append(
            {
                "label": str(item.get(label_key) or "Sin dato"),
                "secondary": str(item.get(secondary_key) or "") if secondary_key else "",
                "value": value,
                "pct": max(8.0, pct) if value > 0 else 0.0,
            }
        )
    return output


def _bi_monthly_sales_rows(snapshot: dict[str, object]) -> list[dict[str, object]]:
    return list(get_dashboard_sales_dataset(months=6).get("monthly_sales_rows") or [])


def _bi_monthly_margin_rows(snapshot: dict[str, object]) -> list[dict[str, object]]:
    if not snapshot.get("kpis", {}).get("margin_ready"):
        return []
    month_names = {
        "01": "Ene",
        "02": "Feb",
        "03": "Mar",
        "04": "Abr",
        "05": "May",
        "06": "Jun",
        "07": "Jul",
        "08": "Ago",
        "09": "Sep",
        "10": "Oct",
        "11": "Nov",
        "12": "Dic",
    }
    rows = list(snapshot.get("series_mensual") or [])
    max_value = max((abs(Decimal(str(item.get("margen") or 0))) for item in rows), default=Decimal("0"))
    output: list[dict[str, object]] = []
    for item in rows:
        periodo = str(item.get("periodo") or "")
        label = periodo
        if len(periodo) == 7 and "-" in periodo:
            year, month = periodo.split("-", 1)
            label = f"{month_names.get(month, month)} {year[-2:]}"
        margin_value = item.get("margen")
        if margin_value is None:
            continue
        value = Decimal(str(margin_value or 0))
        pct = float((abs(value) / max_value) * Decimal("100")) if max_value > 0 else 0.0
        output.append(
            {
                "label": label,
                "value": value,
                "pct": max(8.0, pct) if value != 0 else 0.0,
                "tone": "success" if value >= 0 else "danger",
            }
        )
    return output


def _bi_comparison_bar_rows(
    raw_rows: list[dict[str, object]],
    label_key: str,
    amount_key: str = "amount",
    secondary_key: str | None = None,
    limit: int = 6,
) -> list[dict[str, object]]:
    rows = list(raw_rows or [])[:limit]
    max_delta = max((abs(Decimal(str(item.get("delta_pct") or 0))) for item in rows), default=Decimal("0"))
    output: list[dict[str, object]] = []
    for item in rows:
        delta = Decimal(str(item.get("delta_pct") or 0))
        tone = str(item.get("tone") or "warning")
        pct = float((abs(delta) / max_delta) * Decimal("100")) if max_delta > 0 else 0.0
        output.append(
            {
                "label": str(item.get(label_key) or "Sin dato"),
                "secondary": str(item.get(secondary_key) or "") if secondary_key else "",
                "detail": str(item.get("detail") or ""),
                "status": str(item.get("status") or ""),
                "tone": tone,
                "delta_label": f"{delta:.1f}%",
                "value": Decimal(str(item.get(amount_key) or 0)),
                "pct": max(8.0, pct) if delta != 0 else 0.0,
            }
        )
    return output


def _bi_purchase_snapshot() -> dict[str, object]:
    today = timezone.localdate()
    solicitudes_abiertas = SolicitudCompra.objects.exclude(estatus=SolicitudCompra.STATUS_RECHAZADA)
    ordenes_abiertas = OrdenCompra.objects.exclude(estatus=OrdenCompra.STATUS_CERRADA)
    recepciones_abiertas = RecepcionCompra.objects.exclude(estatus=RecepcionCompra.STATUS_CERRADA)
    solicitudes_vencidas = solicitudes_abiertas.filter(fecha_requerida__lt=today).count()
    recepciones_abiertas_count = recepciones_abiertas.count()
    return {
        "solicitudes_abiertas": solicitudes_abiertas.count(),
        "solicitudes_aprobadas": solicitudes_abiertas.filter(estatus=SolicitudCompra.STATUS_APROBADA).count(),
        "solicitudes_vencidas": solicitudes_vencidas,
        "ordenes_abiertas": ordenes_abiertas.count(),
        "ordenes_por_recibir": ordenes_abiertas.filter(
            fecha_entrega_estimada__isnull=False,
            fecha_entrega_estimada__lte=today + timedelta(days=3),
        ).count(),
        "recepciones_abiertas": recepciones_abiertas_count,
        "status": "En seguimiento" if (solicitudes_vencidas or recepciones_abiertas_count) else "Controlado",
        "tone": "warning" if (solicitudes_vencidas or recepciones_abiertas_count) else "success",
        "detail": (
            "Hay documentos de compra que aún requieren cierre."
            if (solicitudes_vencidas or recepciones_abiertas_count)
            else "El flujo documental de compra está controlado."
        ),
    }


def _bi_inventory_snapshot() -> dict[str, object]:
    rows = list(ExistenciaInsumo.objects.select_related("insumo")[:2000])
    total = len(rows)
    criticos = 0
    bajo_reorden = 0
    for row in rows:
        stock = _to_decimal(getattr(row, "stock_actual", 0))
        reorden = _to_decimal(getattr(row, "punto_reorden", 0))
        minimo = _to_decimal(getattr(row, "stock_minimo", 0))
        if stock <= 0:
            criticos += 1
        elif reorden > 0 and stock < reorden:
            bajo_reorden += 1
        elif minimo > 0 and stock < minimo:
            bajo_reorden += 1
    movimientos_hoy = MovimientoInventario.objects.filter(fecha__date=timezone.now().date()).count()
    return {
        "total": total,
        "criticos": criticos,
        "bajo_reorden": bajo_reorden,
        "movimientos_hoy": movimientos_hoy,
        "status": "Con riesgo" if criticos else ("En revisión" if bajo_reorden else "Controlado"),
        "tone": "danger" if criticos else ("warning" if bajo_reorden else "success"),
        "detail": (
            "Hay artículos críticos que ya comprometen surtido."
            if criticos
            else "Hay artículos que piden reabasto."
            if bajo_reorden
            else "El inventario relevante del tablero está bajo control."
        ),
    }


def _bi_production_snapshot() -> dict[str, object]:
    today = timezone.localdate()
    plan_hoy = PlanProduccion.objects.filter(fecha_produccion=today).order_by("-creado_en").first()
    solicitudes_activas = SolicitudVenta.objects.filter(fecha_inicio__lte=today, fecha_fin__gte=today).count()
    planes_abiertos = PlanProduccion.objects.filter(fecha_produccion__gte=today - timedelta(days=7)).exclude(
        estado=PlanProduccion.ESTADO_CERRADO
    ).count()
    return {
        "plan_hoy": plan_hoy.nombre if plan_hoy else "Sin plan para hoy",
        "plan_hoy_estado": plan_hoy.get_estado_display() if plan_hoy else "Sin plan",
        "planes_abiertos": planes_abiertos,
        "solicitudes_activas": solicitudes_activas,
        "status": "En curso" if plan_hoy else "Sin plan",
        "tone": "warning" if not plan_hoy else "success",
        "detail": (
            "No hay plan cargado para hoy."
            if not plan_hoy
            else "Producción ya tiene plan y puede compararse contra demanda activa."
        ),
    }


def _bi_production_summary(date_from, date_to) -> dict[str, object]:
    plans = list(
        PlanProduccion.objects.filter(fecha_produccion__gte=date_from, fecha_produccion__lte=date_to).order_by("fecha_produccion", "id")
    )
    plan_ids = [int(plan.id) for plan in plans]
    items = list(
        PlanProduccionItem.objects.filter(plan_id__in=plan_ids).select_related("receta", "plan").order_by("plan__fecha_produccion")
    )
    total_units = Decimal("0")
    total_cost = Decimal("0")
    final_units = Decimal("0")
    final_recipe_ids: set[int] = set()
    produced_by_recipe: dict[int, dict[str, object]] = {}

    for item in items:
        qty = _to_decimal(item.cantidad)
        if qty <= 0:
            continue
        total_units += qty
        total_cost += _to_decimal(item.costo_total_estimado)
        bucket = produced_by_recipe.setdefault(
            int(item.receta_id),
            {
                "label": item.receta.nombre,
                "value": Decimal("0"),
                "cost": Decimal("0"),
            },
        )
        bucket["value"] = _to_decimal(bucket["value"]) + qty
        bucket["cost"] = _to_decimal(bucket["cost"]) + _to_decimal(item.costo_total_estimado)
        if item.receta.tipo == Receta.TIPO_PRODUCTO_FINAL:
            final_units += qty
            final_recipe_ids.add(int(item.receta_id))

    sold_units = Decimal("0")
    if final_recipe_ids:
        sold_units = _to_decimal(
            VentaHistorica.objects.filter(
                receta_id__in=final_recipe_ids,
                fecha__gte=date_from,
                fecha__lte=date_to,
            ).aggregate(total=Sum("cantidad")).get("total")
        )

    coverage_pct = None
    if sold_units > 0:
        coverage_pct = (final_units * Decimal("100")) / sold_units

    status = "Sin producción"
    tone = "warning"
    detail = "No hay renglones de producción capturados en la ventana BI."
    if total_units > 0 and coverage_pct is not None:
        if coverage_pct >= Decimal("90"):
            status = "Cubre venta"
            tone = "success"
        elif coverage_pct >= Decimal("70"):
            status = "Cobertura ajustada"
            tone = "warning"
        else:
            status = "Producción corta"
            tone = "danger"
        detail = f"Producción final {final_units:.1f} u contra {sold_units:.1f} u vendidas en la ventana BI."
    elif total_units > 0:
        status = "Producción sin comparable"
        tone = "warning"
        detail = "Hay producción en la ventana BI, pero no existe venta final comparable suficiente."

    top_products = sorted(
        produced_by_recipe.values(),
        key=lambda row: (_to_decimal(row.get("value")), str(row.get("label") or "")),
        reverse=True,
    )[:6]
    return {
        "period_label": f"{date_from.isoformat()} a {date_to.isoformat()}",
        "total_units": total_units,
        "total_cost": total_cost,
        "plan_count": len(plans),
        "open_plan_count": sum(1 for plan in plans if plan.estado != PlanProduccion.ESTADO_CERRADO),
        "final_units": final_units,
        "sales_units": sold_units,
        "coverage_pct": coverage_pct,
        "status": status,
        "tone": tone,
        "detail": detail,
        "top_products": top_products,
        "conversion_note": "Conversión a enteros equivalentes pendiente de catálogo específico por presentación.",
    }


def _bi_waste_summary(date_from, date_to) -> dict[str, object]:
    days_window = max((date_to - date_from).days + 1, 1)
    prev_end = date_from - timedelta(days=1)
    prev_start = prev_end - timedelta(days=days_window - 1)

    branch_rows_qs = MermaPOS.objects.filter(fecha__gte=date_from, fecha__lte=date_to)
    branch_rows = list(
        branch_rows_qs
        .select_related("receta", "sucursal")
        .order_by("-fecha", "-id")
    )
    prev_branch_units = _to_decimal(
        MermaPOS.objects.filter(fecha__gte=prev_start, fecha__lte=prev_end).aggregate(total=Sum("cantidad")).get("total")
    )
    branch_units = Decimal("0")
    branch_cost_est = Decimal("0")
    branch_cost_covered = 0
    branch_by_sucursal: dict[str, dict[str, object]] = {}
    recipe_cost_map = get_total_cost_map({int(row.receta_id) for row in branch_rows if row.receta_id})
    for row in branch_rows:
        qty = _to_decimal(row.cantidad)
        branch_units += qty
        code = row.sucursal.codigo if row.sucursal_id else "SIN SUCURSAL"
        bucket = branch_by_sucursal.setdefault(
            code,
            {"label": code, "secondary": row.sucursal.nombre if row.sucursal_id else "Sin sucursal", "value": Decimal("0")},
        )
        bucket["value"] = _to_decimal(bucket["value"]) + qty
        if row.receta_id:
            recipe_cost = recipe_cost_map.get(int(row.receta_id))
            if recipe_cost is not None:
                branch_cost_est += qty * _to_decimal(recipe_cost)
                branch_cost_covered += 1

    cedis_rows = list(
        MovimientoInventario.objects.filter(
            fecha__date__gte=date_from,
            fecha__date__lte=date_to,
            tipo=MovimientoInventario.TIPO_CONSUMO,
            referencia__startswith="MERMA|",
        ).select_related("insumo")
    )
    prev_cedis_units = _to_decimal(
        MovimientoInventario.objects.filter(
            fecha__date__gte=prev_start,
            fecha__date__lte=prev_end,
            tipo=MovimientoInventario.TIPO_CONSUMO,
            referencia__startswith="MERMA|",
        ).aggregate(total=Sum("cantidad")).get("total")
    )
    cedis_units = Decimal("0")
    cedis_cost_est = Decimal("0")
    cedis_cost_covered = 0
    cedis_by_insumo: dict[str, dict[str, object]] = {}
    cost_cache: dict[int, Decimal | None] = {}
    for row in cedis_rows:
        qty = _to_decimal(row.cantidad)
        cedis_units += qty
        label = row.insumo.nombre if row.insumo_id else "Sin insumo"
        bucket = cedis_by_insumo.setdefault(label, {"label": label, "secondary": "Merma CEDIS", "value": Decimal("0")})
        bucket["value"] = _to_decimal(bucket["value"]) + qty
        if row.insumo_id:
            if int(row.insumo_id) not in cost_cache:
                cost_cache[int(row.insumo_id)] = latest_costo_canonico(insumo_id=int(row.insumo_id))
            unit_cost = cost_cache[int(row.insumo_id)]
            if unit_cost is not None:
                cedis_cost_est += qty * _to_decimal(unit_cost)
                cedis_cost_covered += 1

    total_units = branch_units + cedis_units
    prev_total_units = prev_branch_units + prev_cedis_units
    comparison_label = "Base inicial"
    comparison_tone = "warning"
    comparison_detail = "Aún no hay una ventana previa equivalente para merma."
    if prev_total_units > 0:
        delta_pct = ((total_units - prev_total_units) / prev_total_units) * Decimal("100")
        comparison_label = "Sube" if delta_pct >= 0 else "Baja"
        comparison_tone = "warning" if delta_pct >= 0 else "success"
        comparison_detail = f"{abs(delta_pct):.1f}% vs la ventana previa ({prev_start.isoformat()} a {prev_end.isoformat()})"

    return {
        "period_label": f"{date_from.isoformat()} a {date_to.isoformat()}",
        "branch_available": branch_rows_qs.exists(),
        "branch_units": branch_units,
        "branch_cost_est": branch_cost_est,
        "branch_branch_count": len(branch_by_sucursal),
        "branch_cost_note": (
            f"Costo estimado en sucursal sobre {branch_cost_covered} capturas con receta mapeada."
            if branch_cost_covered
            else "Merma sucursal sin costo estimable: faltan recetas mapeadas."
        ),
        "cedis_units": cedis_units,
        "cedis_cost_est": cedis_cost_est,
        "cedis_row_count": len(cedis_rows),
        "cedis_cost_note": (
            f"Costo estimado en CEDIS sobre {cedis_cost_covered} movimientos con costo canónico."
            if cedis_cost_covered
            else "Merma CEDIS sin costo estimable: faltan costos canónicos para esos insumos."
        ),
        "comparison_label": comparison_label,
        "comparison_tone": comparison_tone,
        "comparison_detail": comparison_detail,
        "branch_rows": sorted(branch_by_sucursal.values(), key=lambda row: _to_decimal(row.get("value")), reverse=True)[:6],
        "cedis_rows": sorted(cedis_by_insumo.values(), key=lambda row: _to_decimal(row.get("value")), reverse=True)[:6],
    }


def _bi_forecast_summary(periodo_mes: str) -> dict[str, object]:
    try:
        year, month = periodo_mes.split("-")
        y = int(year)
        m = int(month)
    except Exception:
        today = timezone.localdate()
        y = today.year
        m = today.month
        periodo_mes = f"{y:04d}-{m:02d}"

    pron_rows = list(
        PronosticoVenta.objects.filter(periodo=periodo_mes)
        .values("receta_id", "receta__nombre")
        .annotate(total=Sum("cantidad"))
    )
    plan_rows = list(
        PlanProduccionItem.objects.filter(plan__fecha_produccion__year=y, plan__fecha_produccion__month=m)
        .values("receta_id", "receta__nombre")
        .annotate(total=Sum("cantidad"))
    )

    merged: dict[int, dict[str, object]] = {}
    for row in pron_rows:
        merged[int(row["receta_id"])] = {
            "label": row["receta__nombre"],
            "pronostico": _to_decimal(row["total"]),
            "plan": Decimal("0"),
        }
    for row in plan_rows:
        payload = merged.setdefault(
            int(row["receta_id"]),
            {"label": row["receta__nombre"], "pronostico": Decimal("0"), "plan": Decimal("0")},
        )
        payload["plan"] = _to_decimal(row["total"])

    top_rows: list[dict[str, object]] = []
    recipes_with_gap = 0
    total_forecast = Decimal("0")
    total_plan = Decimal("0")
    for payload in merged.values():
        pron = _to_decimal(payload["pronostico"])
        plan = _to_decimal(payload["plan"])
        delta = plan - pron
        total_forecast += pron
        total_plan += plan
        if delta != 0:
            recipes_with_gap += 1
            top_rows.append(
                {
                    "label": str(payload["label"]),
                    "secondary": f"Plan {plan:.1f} · Forecast {pron:.1f}",
                    "value": abs(delta),
                    "tone": "danger" if delta > 0 else "warning",
                }
            )
    top_rows.sort(key=lambda row: _to_decimal(row["value"]), reverse=True)
    delta_units = total_plan - total_forecast
    deviation_pct = None
    if total_forecast > 0:
        deviation_pct = (abs(delta_units) * Decimal("100")) / total_forecast

    if total_forecast <= 0 and total_plan <= 0:
        status = "Sin datos"
        tone = "warning"
        detail = "No hay forecast ni producción planificada para el periodo."
    elif total_forecast <= 0 and total_plan > 0:
        status = "Rojo"
        tone = "danger"
        detail = "Hay producción planificada sin forecast cargado en el periodo."
    elif deviation_pct is not None and deviation_pct <= Decimal("10"):
        status = "Verde"
        tone = "success"
        detail = "El plan del mes está alineado con el forecast cargado."
    elif deviation_pct is not None and deviation_pct <= Decimal("25"):
        status = "Amarillo"
        tone = "warning"
        detail = "Hay desviación relevante entre plan y forecast."
    else:
        status = "Rojo"
        tone = "danger"
        detail = "La desviación plan vs forecast ya exige ajuste ejecutivo."

    return {
        "period_label": periodo_mes,
        "forecast_units": total_forecast,
        "plan_units": total_plan,
        "delta_units": delta_units,
        "deviation_pct": deviation_pct,
        "status": status,
        "tone": tone,
        "detail": detail,
        "recipes_total": len(merged),
        "recipes_with_gap": recipes_with_gap,
        "top_rows": top_rows[:6],
        "basis_note": "Forecast mensual cargado en ERP. La exclusión automática de semanas atípicas aún no está parametrizada.",
    }


def _bi_operational_plan() -> PlanProduccion | None:
    today = timezone.localdate()
    plan_hoy = PlanProduccion.objects.filter(fecha_produccion=today).order_by("-creado_en").first()
    if plan_hoy:
        return plan_hoy
    return (
        PlanProduccion.objects.exclude(estado=PlanProduccion.ESTADO_CERRADO)
        .filter(fecha_produccion__gte=today)
        .order_by("fecha_produccion", "-creado_en")
        .first()
    )


def _bi_supply_watchlist(limit: int = 6) -> dict[str, object] | None:
    plan = _bi_operational_plan()
    if not plan:
        return None

    items = list(plan.items.select_related("receta")[:250])
    if not items:
        return None

    recipe_ids = [int(item.receta_id) for item in items if getattr(item, "receta_id", None)]
    lineas = list(
        LineaReceta.objects.filter(receta_id__in=recipe_ids, insumo_id__isnull=False)
        .exclude(tipo_linea=LineaReceta.TIPO_SUBSECCION)
        .select_related("insumo", "receta", "insumo__unidad_base")
    )
    if not lineas:
        return None

    lineas_by_recipe: dict[int, list[LineaReceta]] = {}
    canonical_map: dict[int, Insumo] = {}
    canonical_ids: set[int] = set()
    for linea in lineas:
        if not linea.insumo:
            continue
        canonical_map[linea.id] = linea.insumo
        canonical_ids.add(linea.insumo.id)
        lineas_by_recipe.setdefault(int(linea.receta_id), []).append(linea)

    historico_map = {
        int(row["receta_id"]): Decimal(str(row["total"] or 0))
        for row in (
            VentaHistorica.objects.filter(
                receta_id__in=recipe_ids,
                fecha__gte=timezone.localdate() - timedelta(days=45),
            )
            .values("receta_id")
            .annotate(total=Sum("cantidad"))
        )
    }
    existencia_map = {
        int(existencia.insumo_id): existencia
        for existencia in ExistenciaInsumo.objects.filter(insumo_id__in=canonical_ids).select_related("insumo")
    }

    aggregated: dict[int, dict[str, object]] = {}
    for item in items:
        item_qty = Decimal(str(item.cantidad or 0))
        if item_qty <= 0:
            continue
        historico_units = historico_map.get(int(item.receta_id), Decimal("0"))
        for linea in lineas_by_recipe.get(int(item.receta_id), []):
            insumo = canonical_map.get(linea.id)
            if insumo is None:
                continue
            required_qty = Decimal(str(linea.cantidad or 0)) * item_qty
            if required_qty <= 0:
                continue
            bucket = aggregated.setdefault(
                insumo.id,
                {
                    "insumo": insumo,
                    "required_qty": Decimal("0"),
                    "historico_units": Decimal("0"),
                    "recipe_names": [],
                },
            )
            bucket["required_qty"] = Decimal(str(bucket["required_qty"])) + required_qty
            bucket["historico_units"] = Decimal(str(bucket["historico_units"])) + historico_units
            recipe_names = list(bucket["recipe_names"])
            if item.receta.nombre not in recipe_names:
                recipe_names.append(item.receta.nombre)
            bucket["recipe_names"] = recipe_names[:3]

    rows: list[dict[str, object]] = []
    for payload in aggregated.values():
        insumo = payload["insumo"]
        required_qty = Decimal(str(payload["required_qty"] or 0))
        historico_units = Decimal(str(payload["historico_units"] or 0))
        existencia = existencia_map.get(int(insumo.id))
        stock_actual = Decimal(str(getattr(existencia, "stock_actual", 0) or 0))
        shortage = max(required_qty - stock_actual, Decimal("0"))
        readiness = enterprise_readiness_profile(insumo)
        missing = list(readiness.get("missing") or [])
        missing_cost = latest_costo_canonico(insumo_id=insumo.id) is None
        if shortage <= 0 and not missing and not missing_cost:
            continue
        priority_score = (shortage * Decimal("100")) + (Decimal(str(len(missing))) * Decimal("50")) + historico_units
        if missing_cost:
            priority_score += Decimal("25")
        rows.append(
            {
                "insumo_nombre": insumo.nombre,
                "required_qty": required_qty,
                "stock_actual": stock_actual,
                "shortage": shortage,
                "historico_units": historico_units,
                "master_missing": missing,
                "missing_cost": missing_cost,
                "recipe_names": list(payload["recipe_names"] or []),
                "action_url": reverse("maestros:insumo_update", args=[insumo.id]),
                "action_label": "Asegurar artículo",
                "priority_score": priority_score,
            }
        )

    rows.sort(
        key=lambda item: (
            Decimal(str(item.get("priority_score") or 0)),
            Decimal(str(item.get("shortage") or 0)),
            Decimal(str(item.get("historico_units") or 0)),
        ),
        reverse=True,
    )
    if not rows:
        return None
    return {
        "plan_id": plan.id,
        "plan_nombre": plan.nombre,
        "plan_fecha": plan.fecha_produccion,
        "rows": rows[:limit],
        "url": f"{reverse('recetas:plan_produccion')}?plan_id={plan.id}",
        "cta": "Abrir plan",
    }


def _bi_daily_decisions(
    *,
    daily_sales_snapshot: dict[str, object],
    branch_weekday_rows: list[dict[str, object]],
    product_weekday_rows: list[dict[str, object]],
    purchase_snapshot: dict[str, object],
    inventory_snapshot: dict[str, object],
    production_snapshot: dict[str, object],
    waste_summary: dict[str, object],
    forecast_summary: dict[str, object],
    supply_watchlist: dict[str, object] | None,
) -> list[dict[str, object]]:
    decisions: list[dict[str, object]] = []

    def push(priority: int, tone: str, title: str, detail: str, url: str, cta: str) -> None:
        decisions.append(
            {
                "priority": priority,
                "tone": tone,
                "title": title,
                "detail": detail,
                "url": url,
                "cta": cta,
            }
        )

    if int(inventory_snapshot.get("criticos") or 0) > 0:
        push(
            100,
            "danger",
            "Cerrar stock crítico",
            f"Hay {inventory_snapshot.get('criticos', 0)} insumos en crítico y {inventory_snapshot.get('bajo_reorden', 0)} bajo reorden con riesgo directo de surtido.",
            reverse("inventario:alertas"),
            "Abrir alertas",
        )

    if supply_watchlist and list(supply_watchlist.get("rows") or []):
        top_supply = list(supply_watchlist.get("rows") or [])[0]
        missing = list(top_supply.get("master_missing") or [])
        if Decimal(str(top_supply.get("shortage") or 0)) > 0 or missing or bool(top_supply.get("missing_cost")):
            faltante = ", ".join(missing) if missing else ("costo pendiente" if top_supply.get("missing_cost") else "stock corto")
            push(
                98,
                "danger" if Decimal(str(top_supply.get("shortage") or 0)) > 0 else "warning",
                "Asegurar insumo del plan",
                (
                    f"{top_supply.get('insumo_nombre', 'Artículo')} está frenando "
                    f"{supply_watchlist.get('plan_nombre', 'el plan operativo')}: brecha {Decimal(str(top_supply.get('shortage') or 0)):.2f} "
                    f"y faltante {faltante}."
                ),
                str(top_supply.get("action_url") or supply_watchlist.get("url") or reverse("inventario:alertas")),
                str(top_supply.get("action_label") or "Asegurar artículo"),
            )

    if int(purchase_snapshot.get("solicitudes_vencidas") or 0) > 0:
        push(
            95,
            "danger",
            "Liberar solicitudes vencidas",
            f"Hay {purchase_snapshot.get('solicitudes_vencidas', 0)} solicitudes vencidas que ya deberían estar resueltas con compras.",
            reverse("compras:solicitudes"),
            "Abrir compras",
        )

    if str(production_snapshot.get("plan_hoy_estado") or "") == "Sin plan":
        push(
            92,
            "danger",
            "Confirmar producción del día",
            "No hay plan operativo cargado para hoy. No conviene empujar compras o surtido sin ese documento.",
            reverse("recetas:plan_produccion"),
            "Abrir plan",
        )

    if str(forecast_summary.get("status") or "") in {"Rojo", "Amarillo"}:
        deviation_pct = forecast_summary.get("deviation_pct")
        deviation_label = f"{_to_decimal(deviation_pct):.1f}%" if deviation_pct is not None else "sin %"
        push(
            90,
            "danger" if str(forecast_summary.get("status")) == "Rojo" else "warning",
            "Ajustar forecast del periodo",
            f"Forecast {forecast_summary.get('period_label')} en {forecast_summary.get('status')} con desviación {deviation_label}.",
            reverse("recetas:plan_produccion"),
            "Abrir forecast",
        )

    if _to_decimal(waste_summary.get("branch_units")) + _to_decimal(waste_summary.get("cedis_units")) > 0:
        push(
            88,
            "warning" if str(waste_summary.get("comparison_label") or "") == "Sube" else "success",
            "Atender merma operativa",
            (
                f"Merma sucursal {_to_decimal(waste_summary.get('branch_units')):.1f} u "
                f"y CEDIS {_to_decimal(waste_summary.get('cedis_units')):.1f} u. "
                f"{waste_summary.get('comparison_detail')}"
            ),
            reverse("control:discrepancias"),
            "Abrir merma",
        )

    if str(daily_sales_snapshot.get("comparison_label") or "") == "Abajo":
        push(
            80,
            str(daily_sales_snapshot.get("comparison_tone") or "warning"),
            "Revisar caída del corte reciente",
            str(daily_sales_snapshot.get("comparison_detail") or "La venta del corte reciente viene abajo contra la referencia inmediata."),
            reverse("reportes:bi"),
            "Abrir BI",
        )

    top_branch = branch_weekday_rows[0] if branch_weekday_rows else None
    if top_branch:
        push(
            60,
            str(top_branch.get("tone") or "warning"),
            f"Revisar sucursal {top_branch.get('branch_name', 'Sucursal')}",
            str(top_branch.get("detail") or ""),
            reverse("reportes:bi"),
            "Ver BI",
        )

    top_product = product_weekday_rows[0] if product_weekday_rows else None
    if top_product:
        push(
            58,
            str(top_product.get("tone") or "warning"),
            f"Revisar producto {top_product.get('recipe_name', 'Producto')}",
            str(top_product.get("detail") or ""),
            reverse("reportes:bi"),
            "Ver BI",
        )

    if not decisions:
        push(
            10,
            "success",
            "Operación sin alertas dominantes",
            "El corte reciente no muestra excepciones críticas en ventas, stock, compras o producción.",
            reverse("reportes:bi"),
            "Actualizar BI",
        )

    decisions.sort(key=lambda item: int(item.get("priority") or 0), reverse=True)
    return decisions[:5]


def _reportes_enterprise_chain(
    *,
    focus: str,
    open_count: int,
    blocked_count: int,
    date_from: str | None = None,
    date_to: str | None = None,
    nivel: str | None = None,
) -> list[dict[str, object]]:
    def _enrich(items: list[dict[str, object]], owner: str) -> list[dict[str, object]]:
        total = len(items)
        enriched: list[dict[str, object]] = []
        for index, item in enumerate(items):
            dependency = items[index - 1] if index > 0 else None
            completion = int(round(((index + 1) / total) * 100)) if total else 0
            enriched.append(
                {
                    **item,
                    "owner": owner,
                    "next_step": item.get("cta", "Abrir"),
                    "completion": completion,
                    "depends_on": dependency.get("title") if dependency else "Inicio del flujo",
                    "dependency_status": dependency.get("status") if dependency else "Listo",
                }
            )
        return enriched

    if focus == "costeo":
        return _enrich([
            {
                "step": "01",
                "title": "Maestro ERP",
                "count": blocked_count,
                "status": "Bloqueos" if blocked_count else "Listo",
                "tone": "warning" if blocked_count else "success",
                "detail": "Artículos con faltantes maestros o costo incompleto dentro del costeo.",
                "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
                "cta": "Abrir maestro",
            },
            {
                "step": "02",
                "title": "BOM costeado",
                "count": open_count,
                "status": "Con costo" if open_count else "Sin cobertura",
                "tone": "success" if open_count else "warning",
                "detail": "Recetas con cobertura suficiente para costeo consolidado.",
                "url": reverse("reportes:costo_receta"),
                "cta": "Ver costeo",
            },
            {
                "step": "03",
                "title": "Precio sugerido",
                "count": open_count,
                "status": "Analítico",
                "tone": "primary",
                "detail": "Precio objetivo derivado de costo y margen para lectura ejecutiva.",
                "url": reverse("reportes:costo_receta"),
                "cta": "Revisar precios",
            },
        ], "Costeo")
    if focus == "consumo":
        return _enrich([
            {
                "step": "01",
                "title": "Maestro ERP",
                "count": blocked_count,
                "status": "Bloqueos" if blocked_count else "Listo",
                "tone": "warning" if blocked_count else "success",
                "detail": "Artículos con uso operativo pendiente dentro del periodo consultado.",
                "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
                "cta": "Abrir maestro",
            },
            {
                "step": "02",
                "title": "Consumo consolidado",
                "count": open_count,
                "status": "Periodo activo",
                "tone": "primary",
                "detail": f"Movimientos entre {date_from} y {date_to}.",
                "url": reverse("reportes:consumo") + f"?date_from={date_from}&date_to={date_to}&tipo=ALL",
                "cta": "Ver periodo",
            },
            {
                "step": "03",
                "title": "Costo y reposición",
                "count": blocked_count,
                "status": "Por revisar" if blocked_count else "Alineado",
                "tone": "warning" if blocked_count else "success",
                "detail": "Consumibles con impacto en compras o reabasto por maestro incompleto.",
                "url": reverse("reportes:faltantes"),
                "cta": "Ver reabasto",
            },
        ], "Inventario y Compras")
    if focus == "faltantes":
        return _enrich([
            {
                "step": "01",
                "title": "Maestro ERP",
                "count": blocked_count,
                "status": "Bloqueos" if blocked_count else "Listo",
                "tone": "warning" if blocked_count else "success",
                "detail": "Artículos debajo de mínimo con ficha maestra aún incompleta.",
                "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
                "cta": "Abrir maestro",
            },
            {
                "step": "02",
                "title": "Stock crítico",
                "count": open_count,
                "status": "Nivel " + (nivel or "alerta"),
                "tone": "danger" if (nivel or "alerta") in {"critico", "alerta"} else "warning",
                "detail": "Existencias analizadas contra punto de reorden y sugerencia de compra.",
                "url": reverse("reportes:faltantes") + f"?nivel={nivel or 'alerta'}",
                "cta": "Ver alertas",
            },
            {
                "step": "03",
                "title": "Abastecimiento",
                "count": open_count,
                "status": "Accionable",
                "tone": "primary",
                "detail": "Usa este tablero para mover compra, reorden o conciliación de almacén.",
                "url": reverse("compras:solicitudes"),
                "cta": "Ir a compras",
            },
        ], "Abastecimiento")
    return _enrich([
        {
            "step": "01",
            "title": "Fuente ejecutiva",
            "count": open_count,
            "status": "Mensual",
            "tone": "primary",
            "detail": "Serie consolidada de ventas, compras, nómina y logística.",
            "url": reverse("reportes:bi"),
            "cta": "Abrir BI",
        },
        {
            "step": "02",
            "title": "Disciplina maestra",
            "count": blocked_count,
            "status": "Bloqueos" if blocked_count else "Listo",
            "tone": "warning" if blocked_count else "success",
            "detail": "El maestro incompleto afecta consistencia de indicadores y costos.",
            "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
            "cta": "Corregir maestro",
        },
        {
            "step": "03",
            "title": "Acción directiva",
            "count": open_count,
            "status": "Cockpit activo",
            "tone": "primary",
            "detail": "Usa el tablero para seguimiento financiero, comercial y operativo.",
            "url": reverse("dashboard"),
            "cta": "Ir al dashboard",
        },
    ], "Dirección General")


def _reportes_document_stage_rows(
    *,
    focus: str,
    open_count: int,
    blocked_count: int,
    total_count: int,
    date_from: str | None = None,
    date_to: str | None = None,
    nivel: str | None = None,
) -> list[dict[str, object]]:
    if focus == "costeo":
        rows = [
            {
                "label": "Maestro ERP",
                "open": blocked_count,
                "closed": max(total_count - blocked_count, 0),
                "detail": "Recetas soportadas por artículos listos y costos consistentes.",
                "url": reverse("maestros:insumo_list"),
                "owner": "Maestros / Costeo",
                "next_step": "Cerrar faltantes maestros y costo antes del análisis final.",
            },
            {
                "label": "Cobertura de costo",
                "open": open_count,
                "closed": max(total_count - open_count, 0),
                "detail": "Recetas con costo consolidado frente a recetas sin cobertura completa.",
                "url": reverse("reportes:costo_receta"),
                "owner": "Costeo / Finanzas",
                "next_step": "Completar cobertura de costo y validar consolidado por receta.",
            },
            {
                "label": "Lectura ejecutiva",
                "open": open_count,
                "closed": 0,
                "detail": "Precio sugerido y margen objetivo listos para decisión gerencial.",
                "url": reverse("dashboard"),
                "owner": "Dirección General",
                "next_step": "Revisar margen objetivo y tomar decisión comercial.",
            },
        ]
        for row in rows:
            total = int(row["open"]) + int(row["closed"])
            row["completion"] = int(round((int(row["closed"]) / total) * 100)) if total else 100
        return rows
    if focus == "consumo":
        rows = [
            {
                "label": "Maestro ERP",
                "open": blocked_count,
                "closed": max(total_count - blocked_count, 0),
                "detail": "Artículos con ficha lista para análisis.",
                "url": reverse("maestros:insumo_list"),
                "owner": "Maestros / Inventario",
                "next_step": "Completar fichas maestras antes de consolidar consumo.",
            },
            {
                "label": "Movimiento consolidado",
                "open": open_count,
                "closed": 0,
                "detail": f"Consumo entre {date_from} y {date_to}.",
                "url": reverse("reportes:consumo") + f"?date_from={date_from}&date_to={date_to}&tipo=ALL",
                "owner": "Inventario / BI",
                "next_step": "Validar consolidado del periodo y revisar desvíos relevantes.",
            },
            {
                "label": "Acción de reabasto",
                "open": blocked_count,
                "closed": max(total_count - blocked_count, 0),
                "detail": "Cruce de consumo contra gobierno maestro.",
                "url": reverse("reportes:faltantes"),
                "owner": "Compras / Abastecimiento",
                "next_step": "Convertir hallazgos en acción de reabasto y seguimiento.",
            },
        ]
        for row in rows:
            total = int(row["open"]) + int(row["closed"])
            row["completion"] = int(round((int(row["closed"]) / total) * 100)) if total else 100
        return rows
    if focus == "faltantes":
        rows = [
            {
                "label": "Maestro ERP",
                "open": blocked_count,
                "closed": max(total_count - blocked_count, 0),
                "detail": "Artículos listos vs bloqueados para compra.",
                "url": reverse("maestros:insumo_list"),
                "owner": "Maestros / Compras",
                "next_step": "Cerrar faltantes maestros que bloquean abastecimiento.",
            },
            {
                "label": "Existencia crítica",
                "open": open_count,
                "closed": max(total_count - open_count, 0),
                "detail": f"Filtro activo: {nivel or 'alerta'}.",
                "url": reverse("reportes:faltantes") + f"?nivel={nivel or 'alerta'}",
                "owner": "Inventario / Abastecimiento",
                "next_step": "Confirmar criticidad y priorizar reabasto según nivel operativo.",
            },
            {
                "label": "Solicitud de compra",
                "open": open_count,
                "closed": 0,
                "detail": "Sugerencias de compra y reorden accionables.",
                "url": reverse("compras:solicitudes"),
                "owner": "Compras",
                "next_step": "Emitir solicitud u orden de compra según urgencia.",
            },
        ]
        for row in rows:
            total = int(row["open"]) + int(row["closed"])
            row["completion"] = int(round((int(row["closed"]) / total) * 100)) if total else 100
        return rows
    rows = [
        {
            "label": "Serie ejecutiva",
            "open": open_count,
            "closed": 0,
            "detail": "Meses consolidados dentro del tablero BI.",
            "url": reverse("reportes:bi"),
            "owner": "BI / Dirección",
            "next_step": "Mantener serie ejecutiva lista para lectura gerencial.",
        },
        {
            "label": "Maestro ERP",
            "open": blocked_count,
            "closed": max(total_count - blocked_count, 0),
            "detail": "Artículos listos que ya soportan análisis consistente.",
            "url": reverse("maestros:insumo_list"),
            "owner": "Maestros",
            "next_step": "Cerrar brechas de maestro para sostener indicadores limpios.",
        },
        {
            "label": "Seguimiento directivo",
            "open": open_count,
            "closed": 0,
            "detail": "Lectura mensual para decisión operativa.",
            "url": reverse("dashboard"),
            "owner": "Dirección General",
            "next_step": "Usar el cockpit para seguimiento y decisiones ejecutivas.",
        },
    ]
    for row in rows:
        total = int(row["open"]) + int(row["closed"])
        row["completion"] = int(round((int(row["closed"]) / total) * 100)) if total else 100
    return rows


def _reportes_governance_rows(
    rows: list[dict[str, object]],
    owner_default: str = "Reportes / Operación",
) -> list[dict[str, object]]:
    governance_rows: list[dict[str, object]] = []
    for row in rows:
        governance_rows.append(
            {
                "front": row.get("label", "Reportes"),
                "owner": row.get("owner") or owner_default,
                "blockers": int(row.get("open") or 0),
                "completion": int(row.get("completion") or 0),
                "detail": row.get("detail", ""),
                "next_step": row.get("next_step") or "Revisar frente operativo",
                "url": row.get("url") or reverse("reportes:bi"),
                "cta": row.get("cta") or "Abrir",
            }
        )
    return governance_rows


def _reportes_operational_health_cards(
    *,
    focus: str,
    open_count: int,
    blocked_count: int,
    total_count: int,
) -> list[dict[str, object]]:
    ready_count = max(total_count - blocked_count, 0)
    if focus == "costeo":
        return [
            {
                "label": "Recetas con costo",
                "value": open_count,
                "tone": "success" if open_count else "warning",
                "detail": "Recetas con costo consolidado y lectura válida para el periodo.",
            },
            {
                "label": "Bloqueos maestros",
                "value": blocked_count,
                "tone": "warning" if blocked_count else "success",
                "detail": "Faltantes de maestro o costo que degradan el cálculo.",
            },
            {
                "label": "Recetas listas ERP",
                "value": ready_count,
                "tone": "success",
                "detail": "Recetas con estructura y artículos listos para costeo consistente.",
            },
        ]
    if focus == "consumo":
        return [
            {
                "label": "Movimientos consolidados",
                "value": open_count,
                "tone": "primary",
                "detail": "Movimientos dentro del rango activo del reporte.",
            },
            {
                "label": "Maestro bloqueado",
                "value": blocked_count,
                "tone": "warning" if blocked_count else "success",
                "detail": "Artículos con faltantes maestros que afectan costeo o análisis.",
            },
            {
                "label": "Artículos listos ERP",
                "value": ready_count,
                "tone": "success",
                "detail": "Artículos listos para consumo, costo y reposición.",
            },
        ]
    if focus == "faltantes":
        return [
            {
                "label": "Alertas activas",
                "value": open_count,
                "tone": "danger" if open_count else "success",
                "detail": "Artículos por debajo del nivel esperado.",
            },
            {
                "label": "Bloqueos maestros",
                "value": blocked_count,
                "tone": "warning" if blocked_count else "success",
                "detail": "Faltantes de datos maestros que frenan compra o reorden.",
            },
            {
                "label": "Referencias listas",
                "value": ready_count,
                "tone": "success",
                "detail": "Artículos listos para solicitud o reabasto.",
            },
        ]
    return [
        {
            "label": "Meses consolidados",
            "value": open_count,
            "tone": "primary",
            "detail": "Horizonte ejecutivo activo del tablero BI.",
        },
        {
            "label": "Brechas maestras",
            "value": blocked_count,
            "tone": "warning" if blocked_count else "success",
            "detail": "Faltantes del maestro que degradan consistencia de KPIs.",
        },
        {
            "label": "Cobertura ejecutiva",
            "value": total_count,
            "tone": "success",
            "detail": "Indicadores consolidados listos para seguimiento directivo.",
        },
    ]


def _reportes_maturity_summary(*, chain: list[dict], default_url: str) -> dict:
    total_steps = len(chain)
    completed_steps = sum(1 for item in chain if item.get("tone") == "success")
    attention_steps = max(total_steps - completed_steps, 0)
    coverage_pct = int(round((completed_steps / total_steps) * 100)) if total_steps else 0
    next_priority = next((item for item in chain if item.get("tone") != "success"), None)
    return {
        "completed_steps": completed_steps,
        "attention_steps": attention_steps,
        "coverage_pct": coverage_pct,
        "next_priority_title": next_priority.get("title", "Cadena documental estabilizada") if next_priority else "Cadena documental estabilizada",
        "next_priority_detail": next_priority.get("detail", "El tablero ya está alineado con maestro, consumo y acción operativa.") if next_priority else "El tablero ya está alineado con maestro, consumo y acción operativa.",
        "next_priority_url": next_priority.get("url", default_url) if next_priority else default_url,
        "next_priority_cta": next_priority.get("cta", "Abrir tablero") if next_priority else "Abrir tablero",
    }


def _reportes_critical_path_rows(chain: list[dict[str, object]]) -> list[dict[str, object]]:
    severity_order = {"danger": 0, "warning": 1, "success": 2, "primary": 3}
    ranked = sorted(
        chain,
        key=lambda item: (
            severity_order.get(str(item.get("tone") or "warning"), 9),
            -int(item.get("count") or 0),
            int(item.get("completion") or 0),
        ),
    )
    rows: list[dict[str, object]] = []
    for index, item in enumerate(ranked[:4], start=1):
        rows.append(
            {
                "rank": f"R{index}",
                "title": item.get("title", "Tramo de reportes"),
                "owner": item.get("owner", "Reportes / Operación"),
                "status": item.get("status", "Sin estado"),
                "tone": item.get("tone", "warning"),
                "count": int(item.get("count") or 0),
                "completion": int(item.get("completion") or 0),
                "depends_on": item.get("depends_on", "Inicio del flujo"),
                "dependency_status": item.get("dependency_status", "Sin dependencia registrada"),
                "detail": item.get("detail", ""),
                "next_step": item.get("next_step", "Continuar lectura ejecutiva"),
                "url": item.get("url", reverse("reportes:bi")),
                "cta": item.get("cta", "Abrir"),
            }
        )
    return rows


def _reportes_executive_radar_rows(
    governance_rows: list[dict[str, object]],
    *,
    default_owner: str = "Reportes / Operación",
    fallback_url: str,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for row in governance_rows[:4]:
        completion = int(row.get("completion") or 0)
        blockers = int(row.get("blockers") or 0)
        if blockers <= 0 and completion >= 90:
            tone = "success"
            status = "Controlado"
            dominant_blocker = "Sin bloqueo activo"
        elif completion >= 50:
            tone = "warning"
            status = "En seguimiento"
            dominant_blocker = row.get("detail", "") or "Brecha operativa en seguimiento"
        else:
            tone = "danger"
            status = "Con bloqueo"
            dominant_blocker = row.get("detail", "") or "Bloqueo operativo abierto"
        rows.append(
            {
                "phase": row.get("front", "Frente de reportes"),
                "owner": row.get("owner") or default_owner,
                "status": status,
                "tone": tone,
                "blockers": blockers,
                "progress_pct": completion,
                "dominant_blocker": dominant_blocker,
                "depends_on": row.get("front", "Origen del módulo"),
                "dependency_status": row.get("next_step", "Sin dependencia registrada"),
                "next_step": row.get("next_step", "Abrir frente"),
                "url": row.get("url", fallback_url),
                "cta": row.get("cta", "Abrir"),
            }
        )
    return rows


def _reportes_command_center(*, governance_rows: list[dict[str, object]], maturity_summary: dict[str, object], default_owner: str) -> dict[str, object]:
    blockers = sum(int(row.get("blockers") or 0) for row in governance_rows)
    primary_row = max(governance_rows, key=lambda row: int(row.get("blockers") or 0), default={}) if governance_rows else {}
    tone = "success" if blockers == 0 else ("warning" if blockers <= 3 else "danger")
    status = "Listo para operar" if blockers == 0 else ("En atención" if blockers <= 3 else "Crítico")
    return {
        "owner": primary_row.get("owner") or default_owner,
        "status": status,
        "tone": tone,
        "blockers": blockers,
        "next_step": maturity_summary.get("next_priority_detail") or "Continuar cierre documental del módulo.",
        "cta": maturity_summary.get("next_priority_cta") or primary_row.get("cta") or "Abrir",
        "url": maturity_summary.get("next_priority_url") or primary_row.get("url") or reverse("reportes:bi"),
    }


def _reportes_handoff_map(
    *,
    focus: str,
    blocked_count: int,
    open_count: int,
    total_count: int,
    default_url: str,
) -> list[dict]:
    if focus == "costeo":
        return [
            {
                "label": "Maestro -> BOM",
                "detail": "El artículo debe quedar completo y con costo vigente antes de entrar al costeo.",
                "count": blocked_count,
                "tone": "success" if blocked_count == 0 else "warning",
                "status": "Controlado" if blocked_count == 0 else "Bloqueos maestros",
                "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
                "cta": "Abrir maestro",
                "owner": "Maestros / Costeo",
                "depends_on": "Artículo completo + costo vigente",
                "exit_criteria": "El maestro ya sostiene BOM y costo sin brechas activas.",
                "next_step": "Cerrar brechas maestras antes del costeo.",
                "completion": 100 if blocked_count == 0 else 55,
            },
            {
                "label": "BOM -> Costeo",
                "detail": "La estructura de la receta debe cerrar rendimiento y cobertura de componentes.",
                "count": open_count,
                "tone": "success" if open_count else "warning",
                "status": "Cobertura activa" if open_count else "Sin recetas listas",
                "url": default_url,
                "cta": "Ver costeo",
                "owner": "Recetas / Costeo",
                "depends_on": "BOM trazable + recetas listas",
                "exit_criteria": "La cobertura de componentes ya permite costeo consolidado por receta.",
                "next_step": "Completar cobertura de recetas listas.",
                "completion": 100 if open_count else 35,
            },
            {
                "label": "Costeo -> Precio",
                "detail": "El costo consolidado se transforma en precio sugerido para lectura directiva.",
                "count": total_count,
                "tone": "primary",
                "status": "Analítico" if total_count else "Sin base",
                "url": default_url,
                "cta": "Revisar precios",
                "owner": "Costeo / Dirección",
                "depends_on": "Costo consolidado",
                "exit_criteria": "El precio sugerido ya queda disponible para lectura ejecutiva y decisión comercial.",
                "next_step": "Revisar margen y precio sugerido.",
                "completion": 100 if total_count else 20,
            },
        ]
    if focus == "consumo":
        return [
            {
                "label": "Maestro -> Consumo",
                "detail": "El maestro ERP debe estar completo antes de usar el análisis de consumo para costeo y compra.",
                "count": blocked_count,
                "tone": "success" if blocked_count == 0 else "warning",
                "status": "Controlado" if blocked_count == 0 else "Bloqueos maestros",
                "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
                "cta": "Abrir maestro",
                "owner": "Maestros / Inventario",
                "depends_on": "Artículo completo",
                "exit_criteria": "El consumo ya puede leerse sin distorsiones del maestro.",
                "next_step": "Cerrar faltantes de maestro antes del análisis.",
                "completion": 100 if blocked_count == 0 else 55,
            },
            {
                "label": "Consumo -> Reabasto",
                "detail": "Los movimientos consolidados deben alimentar faltantes, compras y reorden.",
                "count": open_count,
                "tone": "primary",
                "status": "Periodo activo" if open_count else "Sin movimientos",
                "url": default_url,
                "cta": "Ver consumo",
                "owner": "Inventario / Planeación",
                "depends_on": "Movimientos consolidados",
                "exit_criteria": "El periodo ya alimenta reposición y lectura de desviaciones.",
                "next_step": "Validar cobertura del periodo.",
                "completion": 100 if open_count else 35,
            },
            {
                "label": "Reabasto -> Compras",
                "detail": "Los artículos listos deben escalar a compra sin fricción documental.",
                "count": max(total_count - blocked_count, 0),
                "tone": "success" if total_count else "warning",
                "status": "Artículos listos" if total_count else "Sin base operativa",
                "url": reverse("compras:solicitudes"),
                "cta": "Ir a compras",
                "owner": "Compras",
                "depends_on": "Consumo validado + maestro estable",
                "exit_criteria": "El reabasto ya puede escalar a documentos de compra sin bloqueo.",
                "next_step": "Documentar abastecimiento.",
                "completion": 100 if total_count else 30,
            },
        ]
    if focus == "faltantes":
        return [
            {
                "label": "Maestro -> Stock",
                "detail": "Sin ficha maestra completa, el reorden queda bloqueado aunque exista alerta de stock.",
                "count": blocked_count,
                "tone": "success" if blocked_count == 0 else "warning",
                "status": "Controlado" if blocked_count == 0 else "Bloqueos maestros",
                "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
                "cta": "Abrir maestro",
                "owner": "Maestros / Inventario",
                "depends_on": "Ficha maestra lista",
                "exit_criteria": "Las alertas ya se apoyan en artículos listos y trazables.",
                "next_step": "Regularizar artículos bajo alerta.",
                "completion": 100 if blocked_count == 0 else 55,
            },
            {
                "label": "Stock -> Reorden",
                "detail": "Las alertas deben convertirse en sugerencia y solicitud accionable.",
                "count": open_count,
                "tone": "danger" if open_count else "success",
                "status": "Alertas activas" if open_count else "Sin alertas",
                "url": default_url,
                "cta": "Ver faltantes",
                "owner": "Inventario / Compras",
                "depends_on": "Existencias + parámetros de reorden",
                "exit_criteria": "Los faltantes ya quedan listos para acción documental.",
                "next_step": "Disparar reorden o compra.",
                "completion": 100 if open_count == 0 else 35,
            },
            {
                "label": "Reorden -> Compras",
                "detail": "El abastecimiento debe pasar a compras con datos consistentes del artículo.",
                "count": max(total_count - blocked_count, 0),
                "tone": "success" if total_count else "warning",
                "status": "Listo para compra" if total_count else "Sin artículos críticos",
                "url": reverse("compras:solicitudes"),
                "cta": "Ir a compras",
                "owner": "Compras",
                "depends_on": "Alerta consolidada",
                "exit_criteria": "Las alertas ya se convierten en abastecimiento o recepción.",
                "next_step": "Emitir solicitud u orden.",
                "completion": 100 if total_count else 25,
            },
        ]
    return [
        {
            "label": "Fuente -> KPIs",
            "detail": "Las series ejecutivas deben conservar consistencia antes de llegar al dashboard directivo.",
            "count": open_count,
            "tone": "primary",
            "status": "Serie activa" if open_count else "Sin serie",
            "url": default_url,
            "cta": "Abrir BI",
            "owner": "BI / Dirección",
            "depends_on": "Fuentes consolidadas",
            "exit_criteria": "La serie ejecutiva ya queda consolidada para lectura directiva.",
            "next_step": "Validar tablero BI mensual.",
            "completion": 100 if open_count else 25,
        },
        {
            "label": "Maestro -> BI",
            "detail": "Las brechas del maestro degradan consistencia de indicadores y lectura ejecutiva.",
            "count": blocked_count,
            "tone": "success" if blocked_count == 0 else "warning",
            "status": "Controlado" if blocked_count == 0 else "Bloqueos maestros",
            "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
            "cta": "Corregir maestro",
            "owner": "Maestros / Dirección",
            "depends_on": "Catálogo ERP estable",
            "exit_criteria": "Los indicadores ya corren sobre datos maestros consistentes.",
            "next_step": "Cerrar brechas del maestro.",
            "completion": 100 if blocked_count == 0 else 55,
        },
        {
            "label": "BI -> Dirección",
            "detail": "La lectura consolidada debe cerrar el ciclo de decisión directiva.",
            "count": total_count,
            "tone": "success" if total_count else "warning",
            "status": "Cockpit activo" if total_count else "Sin indicadores",
            "url": reverse("dashboard"),
            "cta": "Ir al dashboard",
            "owner": "Dirección General",
            "depends_on": "BI + disciplina maestra",
            "exit_criteria": "El tablero ya soporta decisiones ejecutivas con datos confiables.",
            "next_step": "Tomar decisión ejecutiva sobre el periodo.",
            "completion": 100 if total_count else 25,
        },
    ]


def _reportes_release_gate_rows(
    *,
    focus: str,
    open_count: int,
    blocked_count: int,
    total_count: int,
    date_from: str | None = None,
    date_to: str | None = None,
    nivel: str | None = None,
) -> list[dict[str, object]]:
    ready_count = max(total_count - blocked_count, 0)
    if focus == "costeo":
        return [
            {
                "label": "Maestro ERP completo",
                "open": blocked_count,
                "closed": ready_count,
                "detail": "Artículos con costo y ficha maestra suficientes para liberar costeo.",
                "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
            },
            {
                "label": "Cobertura de receta",
                "open": max(total_count - open_count, 0),
                "closed": open_count,
                "detail": "Recetas con costo consolidado y cobertura suficiente de componentes.",
                "url": reverse("reportes:costo_receta"),
            },
            {
                "label": "Lectura ejecutiva",
                "open": 0,
                "closed": total_count,
                "detail": "Precio sugerido y margen disponibles para lectura directiva.",
                "url": reverse("dashboard"),
            },
        ]
    if focus == "consumo":
        return [
            {
                "label": "Maestro ERP completo",
                "open": blocked_count,
                "closed": ready_count,
                "detail": "Artículos listos para análisis de consumo y costo.",
                "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
            },
            {
                "label": "Periodo consolidado",
                "open": 0 if open_count else 1,
                "closed": open_count,
                "detail": f"Consumo consolidado entre {date_from} y {date_to}.",
                "url": reverse("reportes:consumo") + f"?date_from={date_from}&date_to={date_to}&tipo=ALL",
            },
            {
                "label": "Acción de reabasto",
                "open": blocked_count,
                "closed": ready_count,
                "detail": "Consumo ya listo para escalar a faltantes, reorden y compras.",
                "url": reverse("reportes:faltantes"),
            },
        ]
    if focus == "faltantes":
        return [
            {
                "label": "Maestro ERP completo",
                "open": blocked_count,
                "closed": ready_count,
                "detail": "Artículos con ficha suficiente para compra o reorden.",
                "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
            },
            {
                "label": "Alertas accionables",
                "open": 0 if open_count else 1,
                "closed": open_count,
                "detail": f"Nivel activo: {nivel or 'alerta'}.",
                "url": reverse("reportes:faltantes") + f"?nivel={nivel or 'alerta'}",
            },
            {
                "label": "Escalamiento a compras",
                "open": blocked_count,
                "closed": ready_count,
                "detail": "Referencias listas para entrar al flujo documental de compras.",
                "url": reverse("compras:solicitudes"),
            },
        ]
    return [
        {
            "label": "Serie ejecutiva consolidada",
            "open": 0 if open_count else 1,
            "closed": open_count,
            "detail": "Horizonte mensual listo para lectura directiva.",
            "url": reverse("reportes:bi"),
        },
        {
            "label": "Disciplina maestra",
            "open": blocked_count,
            "closed": ready_count,
            "detail": "El maestro no debe degradar consistencia del cockpit ejecutivo.",
            "url": reverse("maestros:insumo_list") + "?erp_status=incompleto",
        },
        {
            "label": "Cockpit directivo",
            "open": 0,
            "closed": total_count,
            "detail": "Indicadores disponibles para seguimiento DG.",
            "url": reverse("dashboard"),
        },
    ]


def _reportes_release_gate_completion(rows: list[dict[str, object]]) -> dict[str, int]:
    total = sum(int(row.get("open", 0)) + int(row.get("closed", 0)) for row in rows)
    closed = sum(int(row.get("closed", 0)) for row in rows)
    pct = int(round((closed / total) * 100)) if total else 0
    return {"closed": closed, "total": total, "pct": pct}


def _enterprise_usage_label(insumo: Insumo) -> str:
    if insumo.tipo_item == Insumo.TIPO_INTERNO:
        return "Producción interna"
    if insumo.tipo_item == Insumo.TIPO_EMPAQUE:
        return "Empaque final"
    return "Compra directa"


def _enterprise_missing_field(missing: list[str]) -> str | None:
    primary_missing = missing[0] if missing else ""
    return (
        "unidad"
        if primary_missing == "unidad base"
        else "proveedor"
        if primary_missing == "proveedor principal"
        else "categoria"
        if primary_missing == "categoría"
        else "codigo_point"
        if primary_missing == "código Point"
        else None
    )


def _build_report_enterprise_meta(insumo: Insumo) -> dict[str, object]:
    profile = enterprise_readiness_profile(insumo)
    missing_field = _enterprise_missing_field(profile["missing"])
    list_query = {"insumo_id": insumo.id, "usage_scope": "reports"}
    if missing_field:
        list_query["missing_field"] = missing_field
    return {
        "enterprise_status": profile["readiness_label"],
        "enterprise_missing": profile["missing"],
        "enterprise_usage_label": _enterprise_usage_label(insumo),
        "enterprise_edit_url": reverse("maestros:insumo_update", args=[insumo.id]),
        "enterprise_list_url": reverse("maestros:insumo_list") + f"?{urlencode(list_query)}",
    }


def _consumo_rows(date_from: str, date_to: str, tipo: str):
    movimientos = MovimientoInventario.objects.select_related("insumo").filter(
        fecha__date__gte=date_from,
        fecha__date__lte=date_to,
    )
    if tipo != "ALL":
        movimientos = movimientos.filter(tipo=tipo)
    member_to_row, _ = _canonical_catalog_maps()
    grouped = {}
    for movimiento in movimientos:
        row = member_to_row.get(movimiento.insumo_id)
        if not row:
            continue
        canonical = row["canonical"]
        bucket = grouped.setdefault(
            canonical.id,
            {
                "insumo_id": canonical.id,
                "insumo__nombre": canonical.nombre,
                "insumo": canonical,
                "cantidad_total": Decimal("0"),
                "ultima_fecha": None,
                "canonical_variant_count": row["variant_count"],
            },
        )
        bucket["cantidad_total"] += _to_decimal(movimiento.cantidad, "0")
        if bucket["ultima_fecha"] is None or (movimiento.fecha and movimiento.fecha > bucket["ultima_fecha"]):
            bucket["ultima_fecha"] = movimiento.fecha

    resumen = sorted(
        grouped.values(),
        key=lambda item: (-_to_decimal(item["cantidad_total"], "0"), item["insumo__nombre"].lower()),
    )
    for item in resumen:
        item.update(_build_report_enterprise_meta(item["insumo"]))
    return movimientos, resumen


def _export_consumo_csv(rows, date_from: str, date_to: str, tipo: str) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="reporte_consumo_{date_from}_{date_to}.csv"'
    writer = csv.writer(response)
    writer.writerow(["Insumo", "Cantidad total", "Ultimo movimiento", "Filtro tipo", "Desde", "Hasta"])
    for row in rows:
        writer.writerow(
            [
                row["insumo__nombre"],
                row["cantidad_total"],
                row["ultima_fecha"].strftime("%Y-%m-%d %H:%M") if row["ultima_fecha"] else "",
                tipo,
                date_from,
                date_to,
            ]
        )
    return response


def _export_consumo_xlsx(rows, date_from: str, date_to: str, tipo: str) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Consumo"
    ws.append(["Insumo", "Cantidad total", "Ultimo movimiento", "Filtro tipo", "Desde", "Hasta"])
    for row in rows:
        ws.append(
            [
                row["insumo__nombre"],
                float(row["cantidad_total"] or 0),
                row["ultima_fecha"].strftime("%Y-%m-%d %H:%M") if row["ultima_fecha"] else "",
                tipo,
                date_from,
                date_to,
            ]
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="reporte_consumo_{date_from}_{date_to}.xlsx"'
    return response


def _faltantes_rows(nivel: str):
    member_to_row, canonical_by_id = _canonical_catalog_maps()
    raw_existencias = list(ExistenciaInsumo.objects.select_related("insumo", "insumo__unidad_base").order_by("insumo__nombre")[:1000])
    grouped = {}
    for existencia in raw_existencias:
        row = member_to_row.get(existencia.insumo_id)
        if not row:
            continue
        canonical = row["canonical"]
        bucket = grouped.get(canonical.id)
        if bucket is None:
            bucket = SimpleNamespace(
                insumo=canonical,
                stock_actual=Decimal("0"),
                stock_minimo=_to_decimal(existencia.stock_minimo, "0"),
                stock_maximo=_to_decimal(existencia.stock_maximo, "0"),
                punto_reorden=_to_decimal(existencia.punto_reorden, "0"),
                inventario_promedio=_to_decimal(existencia.inventario_promedio, "0"),
                dias_llegada_pedido=int(existencia.dias_llegada_pedido or 0),
                consumo_diario_promedio=_to_decimal(existencia.consumo_diario_promedio, "0"),
                canonical_variant_count=canonical_by_id[canonical.id]["variant_count"],
            )
            grouped[canonical.id] = bucket
        bucket.stock_actual += _to_decimal(existencia.stock_actual, "0")

    existencias = list(grouped.values())

    criticos_count = 0
    bajo_count = 0
    rows = []
    for e in existencias:
        stock = e.stock_actual
        reorden = e.punto_reorden
        if stock <= 0:
            e.criticidad = "Alta"
            e.criticidad_badge = "bg-danger"
            e.nivel = "critico"
            criticos_count += 1
        elif stock < reorden:
            e.criticidad = "Media"
            e.criticidad_badge = "bg-warning"
            e.nivel = "bajo"
            bajo_count += 1
        else:
            e.criticidad = "Sin riesgo"
            e.criticidad_badge = "bg-success"
            e.nivel = "ok"

        e.sugerencia_compra = max(reorden - stock, 0)

        include = False
        if nivel == "all":
            include = True
        elif nivel == "alerta":
            include = e.nivel in {"critico", "bajo"}
        else:
            include = e.nivel == nivel

        if include:
            meta = _build_report_enterprise_meta(e.insumo)
            e.enterprise_status = meta["enterprise_status"]
            e.enterprise_missing = meta["enterprise_missing"]
            e.enterprise_usage_label = meta["enterprise_usage_label"]
            e.enterprise_edit_url = meta["enterprise_edit_url"]
            e.enterprise_list_url = meta["enterprise_list_url"]
            rows.append(e)

    return rows, criticos_count, bajo_count


def _export_faltantes_csv(rows, nivel: str) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="reporte_faltantes.csv"'
    writer = csv.writer(response)
    writer.writerow(["Insumo", "Unidad", "Stock actual", "Punto reorden", "Sugerencia compra", "Criticidad", "Nivel filtro"])
    for row in rows:
        writer.writerow(
            [
                row.insumo.nombre,
                row.insumo.unidad_base.codigo if row.insumo.unidad_base else "-",
                row.stock_actual,
                row.punto_reorden,
                row.sugerencia_compra,
                row.criticidad,
                nivel,
            ]
        )
    return response


def _export_faltantes_xlsx(rows, nivel: str) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Faltantes"
    ws.append(["Insumo", "Unidad", "Stock actual", "Punto reorden", "Sugerencia compra", "Criticidad", "Nivel filtro"])
    for row in rows:
        ws.append(
            [
                row.insumo.nombre,
                row.insumo.unidad_base.codigo if row.insumo.unidad_base else "-",
                float(row.stock_actual or 0),
                float(row.punto_reorden or 0),
                float(row.sugerencia_compra or 0),
                row.criticidad,
                nivel,
            ]
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_faltantes.xlsx"'
    return response


def _safe_div(numerator: Decimal, denominator: Decimal) -> Decimal | None:
    if denominator == 0:
        return None
    return numerator / denominator


def _safe_pct_local(delta: Decimal, base: Decimal) -> Decimal | None:
    ratio = _safe_div(delta, base)
    if ratio is None:
        return None
    return ratio * Decimal("100")


def _financial_cost_signal(cost_pct: Decimal | None) -> dict[str, str | None]:
    if cost_pct is None:
        return {"tone": "neutral", "label": "Sin lectura", "detail": None}
    if cost_pct > Decimal("50"):
        return {"tone": "critical", "label": "Accionar pronto", "detail": "Costo MP arriba de 50% sobre venta."}
    if cost_pct > Decimal("45"):
        return {"tone": "danger", "label": "Rojo", "detail": "Costo MP entre 45% y 50% sobre venta."}
    if cost_pct > Decimal("40"):
        return {"tone": "warning", "label": "Amarillo", "detail": "Costo MP entre 40% y 45% sobre venta."}
    if cost_pct >= Decimal("35"):
        return {"tone": "success", "label": "Verde", "detail": "Costo MP entre 35% y 40% sobre venta."}
    return {"tone": "success", "label": "Óptimo", "detail": "Costo MP por debajo de 35% sobre venta."}


def _financial_signal_rank(tone: str | None) -> int:
    order = {"critical": 0, "danger": 1, "warning": 2, "success": 3, "neutral": 4}
    return order.get(str(tone or "neutral"), 4)


def _pricing_priority(*, tone: str | None, gap_amount: Decimal, bucket: str | None) -> dict[str, object]:
    if tone == "critical" or (tone == "danger" and gap_amount < 0):
        return {"label": "Urgente", "rank": 0}
    if tone == "danger" or (tone == "warning" and gap_amount < 0):
        return {"label": "Corrección", "rank": 1}
    if str(bucket or "") == "Promocionar":
        return {"label": "Táctico", "rank": 2}
    return {"label": "Monitorear", "rank": 3}


def _pricing_dg_action(*, tone: str | None, gap_amount: Decimal, bucket: str | None) -> dict[str, object]:
    bucket_value = str(bucket or "")
    if tone == "critical" or (tone == "danger" and gap_amount < 0):
        return {"label": "Subir precio", "tone": "danger", "rank": 0}
    if bucket_value == "Ajustar margen" or (tone == "warning" and gap_amount < 0):
        return {"label": "Corregir costo", "tone": "warning", "rank": 1}
    if bucket_value == "Promocionar":
        return {"label": "Promocionar", "tone": "warning", "rank": 2}
    if bucket_value == "Defender":
        return {"label": "Defender", "tone": "success", "rank": 3}
    return {"label": "Reformular", "tone": "danger", "rank": 4}


def _pricing_dg_action_note(label: str) -> str:
    notes = {
        "Subir precio": "Costo muy presionado contra venta. Requiere ajuste comercial prioritario.",
        "Corregir costo": "El producto aguanta operación, pero el costo MP necesita corrección rápida.",
        "Promocionar": "Tiene margen sano. Se puede empujar volumen sin comprometer rentabilidad MP.",
        "Defender": "Mantener precio y vigilar ejecución. Es parte del portafolio sano.",
        "Reformular": "La receta o configuración requiere rediseño antes de seguir empujando volumen.",
    }
    return notes.get(label, "Monitorear y revisar el siguiente corte.")


def _financial_week_values(limit: int = 8) -> list[date]:
    return sorted(
        RecetaCostoSemanal.objects.filter(
            scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
            receta__tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        .order_by("-week_start")
        .values_list("week_start", flat=True)
        .distinct()[:limit]
    )


def _build_financial_cost_context(
    *,
    margen_pct: Decimal,
    recipe_scope: str = "final",
    familia: str | None = None,
    categoria: str | None = None,
    bucket: str | None = None,
    coverage: str | None = None,
    q: str = "",
) -> dict[str, object]:
    latest_sales_date = _sales_source_context()["latest_date"] or (timezone.localdate() - timedelta(days=1))
    q = (q or "").strip()
    valid_recipe_scopes = {"final", "base"}
    if recipe_scope not in valid_recipe_scopes:
        recipe_scope = "final"
    recipe_tipo = Receta.TIPO_PRODUCTO_FINAL if recipe_scope == "final" else Receta.TIPO_PREPARACION
    supports_sales_lens = recipe_scope == "final"
    valid_coverages = {"complete", "partial", "blocked"}
    if coverage not in valid_coverages:
        coverage = None
    latest_week = (
        RecetaCostoSemanal.objects.filter(
            scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
            receta__tipo=recipe_tipo,
        )
        .aggregate(v=Max("week_start"))
        .get("v")
    )
    previous_week = None
    current_week_rows: list[RecetaCostoSemanal] = []
    trend_rows: list[dict[str, object]] = []
    family_rows: list[dict[str, object]] = []
    increase_rows: list[dict[str, object]] = []
    decrease_rows: list[dict[str, object]] = []
    price_gap_rows: list[dict[str, object]] = []
    bucket_rows: list[dict[str, object]] = []
    total_weekly_cost = Decimal("0")
    total_weekly_mp = Decimal("0")
    total_delta = Decimal("0")
    previous_total = Decimal("0")
    costed_scope_count = 0
    family_options: list[str] = []
    category_options: list[str] = []
    bucket_options: list[str] = ["Defender", "Promocionar", "Ajustar margen", "Revisar portafolio"]
    scope_options = [
        {"value": "final", "label": "Producto final"},
        {"value": "base", "label": "Base interna"},
    ]
    coverage_options = [
        {"value": "", "label": "Toda la red costeadora"},
        {"value": "complete", "label": "Cobertura completa"},
        {"value": "partial", "label": "Cobertura parcial"},
        {"value": "blocked", "label": "Bloqueadas"},
    ]

    week_values = _financial_week_values()
    if latest_week:
        option_qs = RecetaCostoSemanal.objects.filter(
            week_start=latest_week,
            scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
            receta__tipo=recipe_tipo,
        )
        if q:
            option_qs = option_qs.filter(Q(label__icontains=q) | Q(receta__nombre__icontains=q))
        family_options = list(
            option_qs.exclude(familia__isnull=True)
            .exclude(familia__exact="")
            .order_by("familia")
            .values_list("familia", flat=True)
            .distinct()
        )
        category_option_qs = option_qs
        if familia:
            category_option_qs = category_option_qs.filter(familia=familia)
        category_options = list(
            category_option_qs.exclude(categoria__isnull=True)
            .exclude(categoria__exact="")
            .order_by("categoria")
            .values_list("categoria", flat=True)
            .distinct()
        )
    if supports_sales_lens:
        profitability_panel = build_profitability_panel(
            latest_date=latest_sales_date,
            familia=familia,
            categoria=categoria,
            bucket=bucket,
            q=q,
        )
    else:
        profitability_panel = {
            "rows": [],
            "promo_candidates": [],
            "latest_date": latest_sales_date,
            "source_label": "Base interna sin lectura comercial",
            "note": "Las bases internas no se evalúan contra venta final o ASP.",
        }

    recetas_qs = Receta.objects.prefetch_related("lineas", "lineas__insumo").order_by("nombre")
    recetas_qs = recetas_qs.filter(tipo=recipe_tipo)
    if familia:
        recetas_qs = recetas_qs.filter(familia=familia)
    if categoria:
        recetas_qs = recetas_qs.filter(categoria=categoria)
    if q:
        recetas_qs = recetas_qs.filter(nombre__icontains=q)
    recetas = list(recetas_qs[:500])
    member_to_row, _ = _canonical_catalog_maps()
    latest_cost_by_insumo: dict[int, Decimal] = {}
    total_qty_by_canonical: dict[int, Decimal] = {}
    for receta in recetas:
        for linea in receta.lineas.all():
            if not linea.insumo_id:
                continue
            row = member_to_row.get(linea.insumo_id)
            if not row:
                continue
            canonical = row["canonical"]
            total_qty_by_canonical[canonical.id] = total_qty_by_canonical.get(canonical.id, Decimal("0")) + _to_decimal(linea.cantidad, "0")
    for canonical_id in total_qty_by_canonical:
        latest_cost = latest_costo_canonico(insumo_id=canonical_id)
        if latest_cost is not None:
            latest_cost_by_insumo[canonical_id] = _to_decimal(latest_cost, "0")

    recipe_rows: list[dict[str, object]] = []
    coverage_recipe_ids: dict[str, set[int]] = defaultdict(set)
    for receta in recetas:
        costo_total = Decimal("0")
        lineas_total = 0
        lineas_costeadas = 0
        recipe_blocked = False
        for linea in receta.lineas.all():
            if linea.match_status == LineaReceta.STATUS_REJECTED:
                continue
            if linea.tipo_linea == LineaReceta.TIPO_SUBSECCION:
                continue
            lineas_total += 1
            costo_linea = Decimal("0")
            if linea.costo_linea_excel is not None:
                costo_linea = _to_decimal(linea.costo_linea_excel, "0")
            elif linea.cantidad is not None and linea.costo_unitario_snapshot is not None:
                costo_linea = _to_decimal(linea.cantidad, "0") * _to_decimal(linea.costo_unitario_snapshot, "0")
            elif linea.cantidad is not None and linea.insumo_id:
                row = member_to_row.get(linea.insumo_id)
                canonical = row["canonical"] if row else None
                if canonical and canonical.id in latest_cost_by_insumo:
                    costo_linea = _to_decimal(linea.cantidad, "0") * latest_cost_by_insumo[canonical.id]
                if canonical:
                    profile = enterprise_readiness_profile(canonical)
                    if profile["readiness_label"] != "Listo ERP":
                        recipe_blocked = True
                else:
                    recipe_blocked = True
            elif linea.insumo_texto:
                recipe_blocked = True
            if costo_linea > 0:
                lineas_costeadas += 1
                costo_total += costo_linea

        cobertura_pct = (Decimal("100") * Decimal(lineas_costeadas) / Decimal(lineas_total)) if lineas_total else Decimal("0")
        if lineas_total and lineas_costeadas == lineas_total and not recipe_blocked:
            coverage_state = "complete"
        elif recipe_blocked:
            coverage_state = "blocked"
        else:
            coverage_state = "partial"
        recipe_rows.append(
            {
                "receta": receta,
                "costo_total": costo_total.quantize(Decimal("0.01")),
                "lineas_total": lineas_total,
                "lineas_costeadas": lineas_costeadas,
                "cobertura_pct": cobertura_pct.quantize(Decimal("0.01")),
                "blocked": recipe_blocked,
                "coverage_state": coverage_state,
            }
        )
        if lineas_total:
            coverage_recipe_ids[coverage_state].add(receta.id)

    actionable_recipe_rows = [row for row in recipe_rows if row["lineas_total"] > 0]
    if coverage:
        actionable_recipe_rows = [row for row in actionable_recipe_rows if row["coverage_state"] == coverage]
    complete_rows = [
        row for row in actionable_recipe_rows if row["lineas_costeadas"] == row["lineas_total"] and not row["blocked"]
    ]
    pending_rows = sorted(
        [row for row in actionable_recipe_rows if row not in complete_rows],
        key=lambda item: (item["cobertura_pct"], item["lineas_total"] - item["lineas_costeadas"], item["receta"].nombre),
    )[:10]
    blocked_recipes = sum(1 for row in actionable_recipe_rows if row["blocked"])

    profitability_rows = list(profitability_panel.get("rows", []))
    promo_candidates = list(profitability_panel.get("promo_candidates", []))
    coverage_recipe_id_set = coverage_recipe_ids.get(coverage, set()) if coverage else set()
    if coverage:
        profitability_rows = [
            row for row in profitability_rows if int(row.get("receta_id") or 0) in coverage_recipe_id_set
        ]
        promo_candidates = [
            row for row in promo_candidates if int(row.get("receta_id") or 0) in coverage_recipe_id_set
        ]
    profitability_panel = {
        **profitability_panel,
        "rows": profitability_rows[:18],
        "promo_candidates": promo_candidates[:4],
    }

    allowed_recipe_ids: set[int] | None = None
    profitability_recipe_ids = {int(item["receta_id"]) for item in profitability_rows if item.get("receta_id")}
    if bucket:
        allowed_recipe_ids = set(profitability_recipe_ids)
    if coverage:
        allowed_recipe_ids = (allowed_recipe_ids & coverage_recipe_id_set) if allowed_recipe_ids is not None else set(coverage_recipe_id_set)

    if latest_week:
        for week in week_values:
            weekly_qs = RecetaCostoSemanal.objects.filter(
                week_start=week,
                scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
                receta__tipo=recipe_tipo,
            )
            if familia:
                weekly_qs = weekly_qs.filter(familia=familia)
            if categoria:
                weekly_qs = weekly_qs.filter(categoria=categoria)
            if q:
                weekly_qs = weekly_qs.filter(Q(label__icontains=q) | Q(receta__nombre__icontains=q))
            if allowed_recipe_ids is not None:
                if allowed_recipe_ids:
                    weekly_qs = weekly_qs.filter(receta_id__in=list(allowed_recipe_ids))
                else:
                    weekly_qs = weekly_qs.none()
            qs = list(weekly_qs.select_related("receta"))
            week_cost = sum((_to_decimal(item.costo_total, "0") for item in qs), Decimal("0"))
            week_mp = sum((_to_decimal(item.costo_mp, "0") for item in qs), Decimal("0"))
            trend_rows.append(
                {
                    "label": week.strftime("%Y-%m-%d"),
                    "costo_total": week_cost.quantize(Decimal("0.01")),
                    "costo_mp": week_mp.quantize(Decimal("0.01")),
                    "count": len(qs),
                }
            )

        current_week_qs = RecetaCostoSemanal.objects.filter(
            week_start=latest_week,
            scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
            receta__tipo=recipe_tipo,
        )
        if familia:
            current_week_qs = current_week_qs.filter(familia=familia)
        if categoria:
            current_week_qs = current_week_qs.filter(categoria=categoria)
        if q:
            current_week_qs = current_week_qs.filter(Q(label__icontains=q) | Q(receta__nombre__icontains=q))
        if allowed_recipe_ids is not None:
            if allowed_recipe_ids:
                current_week_qs = current_week_qs.filter(receta_id__in=list(allowed_recipe_ids))
            else:
                current_week_qs = current_week_qs.none()
        current_week_rows = list(current_week_qs.select_related("receta"))
        previous_candidates = [item for item in week_values if item < latest_week]
        previous_week = previous_candidates[-1] if previous_candidates else None
        total_weekly_cost = sum((_to_decimal(item.costo_total, "0") for item in current_week_rows), Decimal("0"))
        total_weekly_mp = sum((_to_decimal(item.costo_mp, "0") for item in current_week_rows), Decimal("0"))
        total_delta = sum(
            (_to_decimal(item.delta_total, "0") for item in current_week_rows if item.delta_total is not None),
            Decimal("0"),
        )
        previous_total = total_weekly_cost - total_delta
        costed_scope_count = sum(1 for item in current_week_rows if _to_decimal(item.costo_total, "0") > 0)

        family_map: dict[str, dict[str, Decimal | str | int]] = defaultdict(
            lambda: {"label": "Sin familia", "costo_total": Decimal("0"), "delta_total": Decimal("0"), "count": 0}
        )
        for item in current_week_rows:
            family_label = (item.familia or "Sin familia").strip() or "Sin familia"
            family_bucket = family_map[family_label]
            family_bucket["label"] = family_label
            family_bucket["costo_total"] = _to_decimal(family_bucket["costo_total"], "0") + _to_decimal(item.costo_total, "0")
            family_bucket["delta_total"] = _to_decimal(family_bucket["delta_total"], "0") + _to_decimal(item.delta_total, "0")
            family_bucket["count"] = int(family_bucket["count"]) + 1
        family_rows = sorted(
            [
                {
                    "label": label,
                    "costo_total": _to_decimal(payload["costo_total"], "0").quantize(Decimal("0.01")),
                    "delta_total": _to_decimal(payload["delta_total"], "0").quantize(Decimal("0.01")),
                    "count": payload["count"],
                }
                for label, payload in family_map.items()
            ],
            key=lambda item: (-_to_decimal(item["costo_total"], "0"), item["label"]),
        )[:8]

        delta_candidates = []
        for item in current_week_rows:
            if item.delta_total is None or _to_decimal(item.delta_total, "0") == 0:
                continue
            delta_candidates.append(
                {
                    "label": item.label,
                    "delta_total": _to_decimal(item.delta_total, "0").quantize(Decimal("0.01")),
                    "delta_pct": (_to_decimal(item.delta_pct, "0").quantize(Decimal("0.01")) if item.delta_pct is not None else None),
                    "costo_total": _to_decimal(item.costo_total, "0").quantize(Decimal("0.01")),
                    "familia": item.familia or "",
                    "categoria": item.categoria or "",
                }
            )
        increase_rows = sorted(
            [row for row in delta_candidates if _to_decimal(row["delta_total"], "0") > 0],
            key=lambda item: (-_to_decimal(item["delta_total"], "0"), item["label"]),
        )[:6]
        decrease_rows = sorted(
            [row for row in delta_candidates if _to_decimal(row["delta_total"], "0") < 0],
            key=lambda item: (_to_decimal(item["delta_total"], "0"), item["label"]),
        )[:6]

        if supports_sales_lens:
            cost_map = _recipe_cost_map_for_sales_lens(
                latest_week=latest_week,
                familia=familia,
                categoria=categoria,
                q=q,
            )
            sales_rows = (
                _active_sales_queryset(start_date=latest_sales_date - timedelta(days=27), end_date=latest_sales_date)
                .filter(total_amount__gt=0, receta_id__in=list(cost_map.keys()))
                .values("receta_id", "receta__nombre")
                .annotate(revenue=Sum("total_amount"), quantity=Sum("quantity"))
            )
            target_margin_ratio = margen_pct / Decimal("100")
            gap_candidates: list[dict[str, object]] = []
            for row in sales_rows:
                receta_id = int(row["receta_id"])
                qty = _to_decimal(row["quantity"], "0")
                revenue = _to_decimal(row["revenue"], "0")
                if qty <= 0 or revenue <= 0:
                    continue
                asp = revenue / qty
                unit_cost = cost_map[receta_id]
                if target_margin_ratio >= Decimal("1"):
                    continue
                suggested_price = unit_cost / (Decimal("1") - target_margin_ratio)
                gap_amount = asp - suggested_price
                gap_pct = _safe_pct_local(gap_amount, suggested_price)
                cost_pct = _safe_pct_local(unit_cost, asp)
                gap_candidates.append(
                    {
                        "receta_id": receta_id,
                        "label": row["receta__nombre"],
                        "asp": asp.quantize(Decimal("0.01")),
                        "unit_cost": unit_cost.quantize(Decimal("0.01")),
                        "cost_pct": cost_pct.quantize(Decimal("0.01")) if cost_pct is not None else None,
                        "suggested_price": suggested_price.quantize(Decimal("0.01")),
                        "gap_amount": gap_amount.quantize(Decimal("0.01")),
                        "gap_pct": gap_pct.quantize(Decimal("0.01")) if gap_pct is not None else None,
                        "revenue": revenue.quantize(Decimal("0.01")),
                        "quantity": qty.quantize(Decimal("0.01")),
                    }
                )
            price_gap_rows = sorted(gap_candidates, key=lambda item: (_to_decimal(item["gap_amount"], "0"), item["label"]))[:8]

    bucket_map: dict[str, int] = defaultdict(int)
    for item in profitability_panel.get("rows", []):
        bucket_map[str(item.get("bucket") or "Sin lectura")] += 1
    bucket_rows = [{"label": label, "count": count} for label, count in bucket_map.items()]
    bucket_rows.sort(key=lambda item: (-item["count"], item["label"]))

    coverage_pct = _safe_pct_local(Decimal(len(complete_rows)), Decimal(len(actionable_recipe_rows))) if actionable_recipe_rows else None
    margin_rows = profitability_panel.get("rows", [])
    avg_margin_pct = None
    avg_cost_pct = profitability_panel.get("avg_cost_pct")
    avg_cost_signal = _financial_cost_signal(avg_cost_pct if isinstance(avg_cost_pct, Decimal) else None)
    if margin_rows:
        margin_values = [_to_decimal(item.get("margin_pct"), "0") for item in margin_rows if item.get("margin_pct") is not None]
        if margin_values:
            avg_margin_pct = (sum(margin_values, Decimal("0")) / Decimal(len(margin_values))).quantize(Decimal("0.01"))
    gap_map = {int(item.get("receta_id") or 0): item for item in price_gap_rows if item.get("receta_id")}
    pricing_action_rows: list[dict[str, object]] = []
    for row in profitability_rows:
        receta_id = int(row.get("receta_id") or 0)
        gap_row = gap_map.get(receta_id)
        if not gap_row:
            continue
        gap_amount = _to_decimal(gap_row.get("gap_amount"), "0")
        visible_impact = max(-gap_amount, Decimal("0")) * _to_decimal(row.get("quantity"), "0")
        priority = _pricing_priority(
            tone=str(row.get("cost_signal_tone") or ""),
            gap_amount=gap_amount,
            bucket=str(row.get("bucket") or ""),
        )
        dg_action = _pricing_dg_action(
            tone=str(row.get("cost_signal_tone") or ""),
            gap_amount=gap_amount,
            bucket=str(row.get("bucket") or ""),
        )
        pricing_action_rows.append(
            {
                "receta_id": receta_id,
                "label": row.get("label"),
                "familia": row.get("familia") or "",
                "categoria": row.get("categoria") or "",
                "bucket": row.get("bucket") or "",
                "recommendation": row.get("recommendation") or "",
                "margin_pct": row.get("margin_pct"),
                "cost_pct": row.get("cost_pct"),
                "cost_signal_tone": row.get("cost_signal_tone"),
                "cost_signal_label": row.get("cost_signal_label"),
                "priority_label": priority["label"],
                "priority_rank": priority["rank"],
                "dg_action_label": dg_action["label"],
                "dg_action_tone": dg_action["tone"],
                "dg_action_rank": dg_action["rank"],
                "dg_action_note": _pricing_dg_action_note(dg_action["label"]),
                "asp": gap_row.get("asp"),
                "unit_cost": gap_row.get("unit_cost"),
                "suggested_price": gap_row.get("suggested_price"),
                "gap_amount": gap_row.get("gap_amount"),
                "gap_pct": gap_row.get("gap_pct"),
                "quantity": row.get("quantity"),
                "visible_impact_amount": visible_impact.quantize(Decimal("0.01")),
                "action_url": _financial_product_action_url(row.get("label"), bucket=row.get("bucket") or None),
            }
        )
    pricing_action_rows.sort(
        key=lambda item: (
            int(item.get("priority_rank") or 99),
            int(item.get("dg_action_rank") or 99),
            -_to_decimal(item.get("visible_impact_amount"), "0"),
            _financial_signal_rank(item.get("cost_signal_tone")),
            _to_decimal(item.get("gap_amount"), "0"),
            -_to_decimal(item.get("quantity"), "0"),
            str(item.get("label") or ""),
        )
    )
    decision_counts = {
        "subir_precio": 0,
        "corregir_costo": 0,
        "promocionar": 0,
        "defender": 0,
        "reformular": 0,
    }
    for item in pricing_action_rows:
        label = str(item.get("dg_action_label") or "")
        key = {
            "Subir precio": "subir_precio",
            "Corregir costo": "corregir_costo",
            "Promocionar": "promocionar",
            "Defender": "defender",
            "Reformular": "reformular",
        }.get(label)
        if key in decision_counts:
            decision_counts[key] += 1
    pricing_total_visible_impact = sum(
        (_to_decimal(item.get("visible_impact_amount"), "0") for item in pricing_action_rows),
        Decimal("0"),
    ).quantize(Decimal("0.01"))
    action_catalog = [
        ("subir_precio", "Subir precio", "danger"),
        ("corregir_costo", "Corregir costo", "warning"),
        ("promocionar", "Promocionar", "warning"),
        ("defender", "Defender", "success"),
        ("reformular", "Reformular", "danger"),
    ]
    pricing_action_summary_rows: list[dict[str, object]] = []
    for key, label, tone in action_catalog:
        rows = [item for item in pricing_action_rows if str(item.get("dg_action_label") or "") == label]
        impact_amount = sum(
            (_to_decimal(item.get("visible_impact_amount"), "0") for item in rows),
            Decimal("0"),
        ).quantize(Decimal("0.01"))
        pricing_action_summary_rows.append(
            {
                "key": key,
                "label": label,
                "tone": tone,
                "count": decision_counts.get(key, 0),
                "impact_amount": impact_amount,
                "top_label": rows[0].get("label") if rows else "",
            }
        )

    return {
        "latest_sales_date": latest_sales_date,
        "latest_week": latest_week,
        "previous_week": previous_week,
        "current_week_rows": current_week_rows,
        "trend_rows": trend_rows,
        "family_rows": family_rows,
        "increase_rows": increase_rows,
        "decrease_rows": decrease_rows,
        "price_gap_rows": price_gap_rows,
        "bucket_rows": bucket_rows,
        "profitability_panel": profitability_panel,
        "recipe_rows": actionable_recipe_rows,
        "pending_rows": pending_rows,
        "complete_rows": complete_rows,
        "blocked_recipes": blocked_recipes,
        "total_recipes": len(actionable_recipe_rows),
        "coverage_pct": coverage_pct.quantize(Decimal("0.01")) if coverage_pct is not None else None,
        "total_weekly_cost": total_weekly_cost.quantize(Decimal("0.01")),
        "total_weekly_mp": total_weekly_mp.quantize(Decimal("0.01")),
        "total_delta": total_delta.quantize(Decimal("0.01")),
        "previous_total": previous_total.quantize(Decimal("0.01")) if previous_total else Decimal("0.00"),
        "costed_scope_count": costed_scope_count,
        "avg_margin_pct": avg_margin_pct,
        "avg_cost_pct": avg_cost_pct.quantize(Decimal("0.01")) if isinstance(avg_cost_pct, Decimal) else avg_cost_pct,
        "avg_cost_signal": avg_cost_signal,
        "pricing_action_rows": pricing_action_rows[:10],
        "pricing_decision_counts": decision_counts,
        "pricing_total_visible_impact": pricing_total_visible_impact,
        "pricing_action_summary_rows": pricing_action_summary_rows,
        "scope_options": scope_options,
        "supports_sales_lens": supports_sales_lens,
        "selected_scope": recipe_scope,
        "family_options": family_options,
        "category_options": category_options,
        "bucket_options": bucket_options,
        "coverage_options": coverage_options,
        "selected_familia": familia or "",
        "selected_categoria": categoria or "",
        "selected_bucket": bucket or "",
        "selected_coverage": coverage or "",
        "selected_q": q,
    }


@login_required
def costo_receta(request: HttpRequest) -> HttpResponse:
    if not can_view_reportes(request.user):
        raise PermissionDenied("No tienes permisos para ver Reportes.")

    margen_pct = _to_decimal(request.GET.get("margen"), "35")
    if margen_pct < 0:
        margen_pct = Decimal("0")
    if margen_pct > 95:
        margen_pct = Decimal("95")
    selected_scope = (request.GET.get("scope") or "final").strip().lower()
    selected_familia = (request.GET.get("familia") or "").strip() or None
    selected_categoria = (request.GET.get("categoria") or "").strip() or None
    selected_bucket = (request.GET.get("bucket") or "").strip() or None
    valid_buckets = {"Defender", "Promocionar", "Ajustar margen", "Revisar portafolio"}
    if selected_bucket not in valid_buckets:
        selected_bucket = None
    selected_coverage = (request.GET.get("coverage") or "").strip() or None
    valid_coverages = {"complete", "partial", "blocked"}
    if selected_coverage not in valid_coverages:
        selected_coverage = None
    selected_q = (request.GET.get("q") or "").strip()
    finance_context = _build_financial_cost_context(
        margen_pct=margen_pct,
        recipe_scope=selected_scope,
        familia=selected_familia,
        categoria=selected_categoria,
        bucket=selected_bucket,
        coverage=selected_coverage,
        q=selected_q,
    )
    profitability_panel = finance_context["profitability_panel"]
    hero_delta_pct = _safe_pct_local(
        _to_decimal(finance_context["total_delta"], "0"),
        _to_decimal(finance_context["previous_total"], "0"),
    )
    context = {
        **finance_context,
        "margen_pct": margen_pct,
        "hero_delta_pct": hero_delta_pct.quantize(Decimal("0.01")) if hero_delta_pct is not None else None,
        "module_tabs": _reportes_module_tabs("financiero"),
    }
    return render(request, "reportes/costo_receta.html", context)


@login_required
def gastos_operativos_captura_manual(request: HttpRequest) -> HttpResponse:
    """Captura manual de un gasto operativo mensual."""
    if not can_view_reportes(request.user):
        raise PermissionDenied

    if request.method == "POST":
        from reportes.models import CategoriaGasto, CentroCosto, GastoOperativoMensual
        import uuid

        try:
            periodo_str = request.POST.get("periodo")
            centro_id = request.POST.get("centro_costo")
            categoria_id = request.POST.get("categoria_gasto")
            monto = request.POST.get("monto")
            tipo_dato = request.POST.get("tipo_dato", "REAL")
            comentario = request.POST.get("comentario", "")

            if not all([periodo_str, centro_id, categoria_id, monto]):
                messages.error(request, "Todos los campos obligatorios deben llenarse.")
                return redirect("reportes:gastos_operativos_importar")

            year, month = map(int, periodo_str.split("-"))
            periodo = date(year, month, 1)
            centro = CentroCosto.objects.get(id=centro_id)
            categoria = CategoriaGasto.objects.get(id=categoria_id)
            monto_decimal = Decimal(monto)

            ext_key = f"MANUAL-{centro.codigo}-{periodo_str}-{categoria.codigo}-{uuid.uuid4().hex[:8]}"

            GastoOperativoMensual.objects.create(
                periodo=periodo,
                centro_costo=centro,
                categoria_gasto=categoria,
                monto=monto_decimal,
                tipo_dato=tipo_dato,
                fuente="MANUAL",
                comentario=comentario,
                capturado_por=request.user,
                external_key=ext_key,
                es_estimado=False,
            )
            messages.success(
                request,
                f"Gasto registrado: {categoria.nombre} · {centro.nombre} · ${monto_decimal:,.2f}",
            )
        except Exception as exc:
            messages.error(request, f"Error al registrar gasto: {exc}")

        return redirect("reportes:gastos_operativos_importar")

    return redirect("reportes:gastos_operativos_importar")


@login_required
def gastos_operativos_importar(request: HttpRequest) -> HttpResponse:
    if not can_view_reportes(request.user):
        raise PermissionDenied("No tienes permisos para ver Reportes.")

    automation_service = OperatingExpenseImportAutomationService()
    target_year = 2026
    if request.method == "POST":
        uploaded_file = request.FILES.get("expense_file")
        if uploaded_file is None:
            messages.error(request, "Selecciona un archivo XLSX antes de intentar la carga.")
            return redirect("reportes:gastos_operativos_importar")
        if not uploaded_file.name.lower().endswith(".xlsx"):
            messages.error(request, "Formato no soportado. Sólo se aceptan archivos .xlsx.")
            return redirect("reportes:gastos_operativos_importar")
        try:
            run = automation_service.process_uploaded_file(
                uploaded_file,
                uploaded_by=request.user,
                target_year=target_year,
            )
        except FileNotFoundError:
            messages.error(request, "No se pudo almacenar el archivo para procesarlo.")
            return redirect("reportes:gastos_operativos_importar")
        if run.status == CargaGastoOperativoArchivo.STATUS_SUCCESS:
            messages.success(
                request,
                (
                    "Carga completada: "
                    f"{run.loaded_rows} registro(s) aplicados, "
                    f"{run.project_refresh_count} proyecto(s) refrescado(s)."
                ),
            )
        elif run.status == CargaGastoOperativoArchivo.STATUS_DUPLICATE:
            messages.warning(request, "El archivo ya había sido procesado anteriormente; se registró como duplicado.")
        else:
            messages.error(
                request,
                "La carga fue rechazada. Revisa el historial y corrige los errores del archivo.",
            )
        return redirect("reportes:gastos_operativos_importar")

    history = list(
        CargaGastoOperativoArchivo.objects.select_related("uploaded_by").order_by("-uploaded_at", "-id")[:20]
    )
    status_counts = {
        "success": CargaGastoOperativoArchivo.objects.filter(
            status=CargaGastoOperativoArchivo.STATUS_SUCCESS
        ).count(),
        "error": CargaGastoOperativoArchivo.objects.filter(
            status=CargaGastoOperativoArchivo.STATUS_ERROR
        ).count(),
        "duplicate": CargaGastoOperativoArchivo.objects.filter(
            status=CargaGastoOperativoArchivo.STATUS_DUPLICATE
        ).count(),
    }
    latest_success = (
        CargaGastoOperativoArchivo.objects.filter(status=CargaGastoOperativoArchivo.STATUS_SUCCESS)
        .order_by("-processed_at", "-id")
        .first()
    )
    latest_error = (
        CargaGastoOperativoArchivo.objects.filter(status=CargaGastoOperativoArchivo.STATUS_ERROR)
        .order_by("-processed_at", "-id")
        .first()
    )
    from reportes.models import CategoriaGasto, CentroCosto
    hoy = date.today()
    periodos: list[str] = []
    year, month = hoy.year, hoy.month
    for _ in range(24):
        periodos.append(f"{year}-{month:02d}")
        month -= 1
        if month == 0:
            month = 12
            year -= 1

    context = {
        "module_tabs": _reportes_module_tabs("gastos_operativos"),
        "history": history,
        "target_year": target_year,
        "status_counts": status_counts,
        "latest_success": latest_success,
        "latest_error": latest_error,
        "accepted_format": "XLSX",
        "upload_help": [
            "El archivo debe contener sucursal, periodo, monto y opcionalmente tipo_dato/categoria_gasto.",
            "Sólo se cargan filas REAL del año 2026; PRESUPUESTO se ignora o rechaza según la política del service.",
            "Cada carga queda trazada con archivo, usuario, errores, sucursales afectadas y proyectos refrescados.",
        ],
        "centros_costo": CentroCosto.objects.filter(tipo="SUCURSAL_VENTA").select_related("sucursal").order_by("nombre"),
        "categorias_gasto": CategoriaGasto.objects.filter(activo=True).order_by("nombre"),
        "periodos_recientes": periodos,
    }
    return render(request, "reportes/gastos_operativos_importar.html", context)


def _budget_upload_history_rows(limit: int = 24) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for audit in BudgetAreaUploadService.history_queryset().order_by("-timestamp", "-id")[:limit]:
        payload = dict(audit.payload or {})
        rows.append(
            {
                "timestamp": audit.timestamp,
                "action": audit.action,
                "user": audit.user,
                "area_key": payload.get("area_key") or "",
                "area_label": payload.get("area_label") or "",
                "expected_filename": payload.get("expected_filename") or "",
                "original_filename": payload.get("original_filename") or "",
                "canonical_filename": payload.get("canonical_filename") or "",
                "status": payload.get("status") or audit.action.removeprefix("BUDGET_UPLOAD_"),
                "detail": payload.get("detail") or "",
                "target_year": payload.get("target_year"),
                "periods": list(payload.get("periods") or []),
                "expected_sheets": list(payload.get("expected_sheets") or []),
                "sheets_imported": list(payload.get("sheets_imported") or []),
                "imports_created": payload.get("imports_created") or 0,
                "imports_updated": payload.get("imports_updated") or 0,
                "lines_created": payload.get("lines_created") or 0,
                "lines_updated": payload.get("lines_updated") or 0,
                "snapshot_rows_created": payload.get("snapshot_rows_created") or 0,
                "snapshot_rows_updated": payload.get("snapshot_rows_updated") or 0,
                "file_hash": payload.get("file_hash") or "",
                "user_label": payload.get("user_label")
                or (audit.user.get_full_name() if audit.user else "")
                or (audit.user.username if audit.user else "Sistema"),
            }
        )
    return rows


def _budget_area_cards(history_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    history_by_area: dict[str, dict[str, object]] = {}
    for row in history_rows:
        area_key = str(row.get("area_key") or "")
        if area_key and area_key not in history_by_area:
            history_by_area[area_key] = row

    import_counts: dict[str, int] = {}
    for import_obj in PresupuestoImport.objects.order_by("id").only("metadata"):
        area_key = str((import_obj.metadata or {}).get("upload_area_key") or "").strip()
        if not area_key:
            continue
        import_counts[area_key] = import_counts.get(area_key, 0) + 1

    cards: list[dict[str, object]] = []
    for definition in BudgetAreaUploadService.list_area_definitions():
        latest = history_by_area.get(definition.key)
        cards.append(
            {
                "definition": definition,
                "latest": latest,
                "import_count": import_counts.get(definition.key, 0),
            }
        )
    return cards


@login_required
def presupuesto_importar_por_area(request: HttpRequest) -> HttpResponse:
    if not can_view_reportes(request.user):
        raise PermissionDenied("No tienes permisos para ver Reportes.")

    service = BudgetAreaUploadService()
    if request.method == "POST":
        area_key = (request.POST.get("area_key") or "").strip()
        uploaded_file = request.FILES.get("budget_file")
        if not area_key:
            messages.error(request, "Selecciona el bloque del área que corresponde al archivo.")
            return redirect("reportes:presupuesto_importar_por_area")
        if uploaded_file is None:
            messages.error(request, "Selecciona un archivo XLSX antes de intentar la carga.")
            return redirect("reportes:presupuesto_importar_por_area")
        try:
            result = service.process_uploaded_file(area_key=area_key, uploaded_file=uploaded_file, uploaded_by=request.user)
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect("reportes:presupuesto_importar_por_area")
        except FileNotFoundError:
            messages.error(request, "No se pudo almacenar el archivo para procesarlo.")
            return redirect("reportes:presupuesto_importar_por_area")
        except Exception as exc:
            messages.error(request, f"La carga fue rechazada: {exc}")
            return redirect("reportes:presupuesto_importar_por_area")

        if result.status == BudgetAreaUploadService.STATUS_DUPLICATE:
            messages.warning(request, f"{result.area_label}: el archivo ya existía y se registró como duplicado.")
        else:
            messages.success(
                request,
                (
                    f"{result.area_label}: {result.lines_created + result.lines_updated} línea(s) procesadas, "
                    f"{result.snapshot_rows_created + result.snapshot_rows_updated} resumen(es) actualizados."
                ),
            )
        return redirect("reportes:presupuesto_importar_por_area")

    history_rows = _budget_upload_history_rows()
    area_cards = _budget_area_cards(history_rows)
    status_counts = {
        "success": sum(1 for row in history_rows if row["status"] == "SUCCESS"),
        "error": sum(1 for row in history_rows if row["status"] == "ERROR"),
        "duplicate": sum(1 for row in history_rows if row["status"] == "DUPLICATE"),
    }
    latest_upload = history_rows[0] if history_rows else None
    context = {
        "module_tabs": _reportes_module_tabs("presupuestos"),
        "area_cards": area_cards,
        "history_rows": history_rows,
        "status_counts": status_counts,
        "latest_upload": latest_upload,
        "upload_help": [
            "Cada bloque acepta sólo el formato ya soportado por su importador existente.",
            "La validación exige hojas esperadas, estructura compatible y evita duplicados por hash.",
            "Cada intento deja bitácora con área, archivo, usuario, periodos y resultado de consolidación.",
        ],
    }
    return render(request, "reportes/presupuesto_importar_por_area.html", context)


@login_required
def consumo(request: HttpRequest) -> HttpResponse:
    if not can_view_reportes(request.user):
        raise PermissionDenied("No tienes permisos para ver Reportes.")
    today = timezone.localdate()
    default_from = today - timedelta(days=30)

    date_from = request.GET.get("date_from") or default_from.isoformat()
    date_to = request.GET.get("date_to") or today.isoformat()
    tipo = (request.GET.get("tipo") or "all").upper()
    valid_tipos = {"ALL", "CONSUMO", "SALIDA", "ENTRADA"}
    if tipo not in valid_tipos:
        tipo = "ALL"

    movimientos, resumen = _consumo_rows(date_from, date_to, tipo)

    export_format = (request.GET.get("export") or "").lower()
    if export_format == "csv":
        return _export_consumo_csv(resumen, date_from, date_to, tipo)
    if export_format == "xlsx":
        return _export_consumo_xlsx(resumen, date_from, date_to, tipo)

    context = {
        "rows": resumen,
        "total_movimientos": movimientos.count(),
        "total_insumos": len(resumen),
        "total_cantidad": sum((row["cantidad_total"] or 0) for row in resumen),
        "governance_ready": sum(1 for row in resumen if row["enterprise_status"] == "Listo ERP"),
        "governance_blocked": sum(1 for row in resumen if row["enterprise_status"] != "Listo ERP"),
        "enterprise_chain": _reportes_enterprise_chain(
            focus="consumo",
            open_count=movimientos.count(),
            blocked_count=sum(1 for row in resumen if row["enterprise_status"] != "Listo ERP"),
            date_from=date_from,
            date_to=date_to,
        ),
        "document_stage_rows": _reportes_document_stage_rows(
            focus="consumo",
            open_count=movimientos.count(),
            blocked_count=sum(1 for row in resumen if row["enterprise_status"] != "Listo ERP"),
            total_count=len(resumen),
            date_from=date_from,
            date_to=date_to,
        ),
        "erp_governance_rows": _reportes_governance_rows(
            _reportes_document_stage_rows(
                focus="consumo",
                open_count=movimientos.count(),
                blocked_count=sum(1 for row in resumen if row["enterprise_status"] != "Listo ERP"),
                total_count=len(resumen),
                date_from=date_from,
                date_to=date_to,
            )
        ),
        "operational_health_cards": _reportes_operational_health_cards(
            focus="consumo",
            open_count=movimientos.count(),
            blocked_count=sum(1 for row in resumen if row["enterprise_status"] != "Listo ERP"),
            total_count=len(resumen),
        ),
        "maturity_summary": _reportes_maturity_summary(
            chain=_reportes_enterprise_chain(
                focus="consumo",
                open_count=movimientos.count(),
                blocked_count=sum(1 for row in resumen if row["enterprise_status"] != "Listo ERP"),
                date_from=date_from,
                date_to=date_to,
            ),
            default_url=reverse("reportes:consumo"),
        ),
        "handoff_map": _reportes_handoff_map(
            focus="consumo",
            blocked_count=sum(1 for row in resumen if row["enterprise_status"] != "Listo ERP"),
            open_count=movimientos.count(),
            total_count=len(resumen),
            default_url=reverse("reportes:consumo") + f"?date_from={date_from}&date_to={date_to}&tipo={tipo}",
        ),
        "release_gate_rows": _reportes_release_gate_rows(
            focus="consumo",
            open_count=movimientos.count(),
            blocked_count=sum(1 for row in resumen if row["enterprise_status"] != "Listo ERP"),
            total_count=len(resumen),
            date_from=date_from,
            date_to=date_to,
        ),
        "filters": {
            "date_from": date_from,
            "date_to": date_to,
            "tipo": tipo,
        },
    }
    context["release_gate_completion"] = _reportes_release_gate_completion(context["release_gate_rows"])
    context["critical_path_rows"] = _reportes_critical_path_rows(context["enterprise_chain"])
    context["executive_radar_rows"] = _reportes_executive_radar_rows(
        context["erp_governance_rows"],
        default_owner="Inventario y Compras",
        fallback_url=reverse("reportes:consumo"),
    )
    context["erp_command_center"] = _reportes_command_center(
        governance_rows=context["erp_governance_rows"],
        maturity_summary=context["maturity_summary"],
        default_owner="Inventario y Compras",
    )
    return render(request, "reportes/consumo.html", context)


@login_required
def faltantes(request: HttpRequest) -> HttpResponse:
    if not can_view_reportes(request.user):
        raise PermissionDenied("No tienes permisos para ver Reportes.")
    nivel = (request.GET.get("nivel") or "alerta").lower()
    valid_levels = {"alerta", "critico", "bajo", "all"}
    if nivel not in valid_levels:
        nivel = "alerta"

    rows, criticos_count, bajo_count = _faltantes_rows(nivel)

    export_format = (request.GET.get("export") or "").lower()
    if export_format == "csv":
        return _export_faltantes_csv(rows, nivel)
    if export_format == "xlsx":
        return _export_faltantes_xlsx(rows, nivel)

    context = {
        "rows": rows,
        "nivel": nivel,
        "criticos_count": criticos_count,
        "bajo_count": bajo_count,
        "alertas_count": criticos_count + bajo_count,
        "governance_ready": sum(1 for row in rows if row.enterprise_status == "Listo ERP"),
        "governance_blocked": sum(1 for row in rows if row.enterprise_status != "Listo ERP"),
        "enterprise_chain": _reportes_enterprise_chain(
            focus="faltantes",
            open_count=criticos_count + bajo_count,
            blocked_count=sum(1 for row in rows if row.enterprise_status != "Listo ERP"),
            nivel=nivel,
        ),
        "document_stage_rows": _reportes_document_stage_rows(
            focus="faltantes",
            open_count=criticos_count + bajo_count,
            blocked_count=sum(1 for row in rows if row.enterprise_status != "Listo ERP"),
            total_count=len(rows),
            nivel=nivel,
        ),
        "erp_governance_rows": _reportes_governance_rows(
            _reportes_document_stage_rows(
                focus="faltantes",
                open_count=criticos_count + bajo_count,
                blocked_count=sum(1 for row in rows if row.enterprise_status != "Listo ERP"),
                total_count=len(rows),
                nivel=nivel,
            )
        ),
        "operational_health_cards": _reportes_operational_health_cards(
            focus="faltantes",
            open_count=criticos_count + bajo_count,
            blocked_count=sum(1 for row in rows if row.enterprise_status != "Listo ERP"),
            total_count=len(rows),
        ),
        "maturity_summary": _reportes_maturity_summary(
            chain=_reportes_enterprise_chain(
                focus="faltantes",
                open_count=criticos_count + bajo_count,
                blocked_count=sum(1 for row in rows if row.enterprise_status != "Listo ERP"),
                nivel=nivel,
            ),
            default_url=reverse("reportes:faltantes"),
        ),
        "handoff_map": _reportes_handoff_map(
            focus="faltantes",
            blocked_count=sum(1 for row in rows if row.enterprise_status != "Listo ERP"),
            open_count=criticos_count + bajo_count,
            total_count=len(rows),
            default_url=reverse("reportes:faltantes") + f"?nivel={nivel}",
        ),
        "release_gate_rows": _reportes_release_gate_rows(
            focus="faltantes",
            open_count=criticos_count + bajo_count,
            blocked_count=sum(1 for row in rows if row.enterprise_status != "Listo ERP"),
            total_count=len(rows),
            nivel=nivel,
        ),
    }
    context["release_gate_completion"] = _reportes_release_gate_completion(context["release_gate_rows"])
    context["critical_path_rows"] = _reportes_critical_path_rows(context["enterprise_chain"])
    context["executive_radar_rows"] = _reportes_executive_radar_rows(
        context["erp_governance_rows"],
        default_owner="Abastecimiento",
        fallback_url=reverse("reportes:faltantes"),
    )
    context["erp_command_center"] = _reportes_command_center(
        governance_rows=context["erp_governance_rows"],
        maturity_summary=context["maturity_summary"],
        default_owner="Abastecimiento",
    )
    return render(request, "reportes/faltantes.html", context)


def _export_bi_csv(snapshot: dict) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="reporte_bi_mensual.csv"'
    writer = csv.writer(response)
    writer.writerow(["Periodo", "Compras", "Ventas", "Nomina", "Margen", "Entregas"])
    for row in snapshot["series_mensual"]:
        writer.writerow(
            [
                row["periodo"],
                row["compras"],
                row["ventas"],
                row["nomina"],
                row["margen"],
                row["entregas"],
            ]
        )
    return response


def _export_bi_xlsx(snapshot: dict) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "BI"
    ws.append(["Periodo", "Compras", "Ventas", "Nomina", "Margen", "Entregas"])
    for row in snapshot["series_mensual"]:
        ws.append(
            [
                row["periodo"],
                float(row["compras"] or 0),
                float(row["ventas"] or 0),
                float(row["nomina"] or 0),
                float(row["margen"] or 0),
                int(row["entregas"] or 0),
            ]
        )
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_bi_mensual.xlsx"'
    return response


def _simple_pdf_bytes(*, title: str, lines: list[str]) -> bytes:
    def _escape(text: str) -> str:
        return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

    page_width = 792
    page_height = 612
    y = 560
    content_lines = ["BT", "/F1 12 Tf", "36 560 Td"]
    first = True
    for raw in [title, *lines[:36]]:
        text = _escape(raw)
        if first:
            content_lines.append(f"({text}) Tj")
            first = False
        else:
            content_lines.append("T*")
            content_lines.append(f"({text}) Tj")
        y -= 14
        if y < 40:
            break
    content_lines.append("ET")
    content = "\n".join(content_lines).encode("latin-1", errors="replace")

    objects = [
        b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj",
        b"2 0 obj << /Type /Pages /Count 1 /Kids [3 0 R] >> endobj",
        f"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width} {page_height}] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj".encode(),
        b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj",
        b"5 0 obj << /Length " + str(len(content)).encode() + b" >> stream\n" + content + b"\nendstream endobj",
    ]

    output = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(len(output))
        output.extend(obj)
        output.extend(b"\n")
    xref_pos = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode())
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode())
    output.extend(
        f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_pos}\n%%EOF".encode()
    )
    return bytes(output)


def _export_branch_bi_csv(branch_panel: dict[str, object], contribution_panel: dict[str, object]) -> HttpResponse:
    branch_slug = str(branch_panel.get("selected_branch_code") or "sucursal").lower()
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="bi_sucursal_{branch_slug}.csv"'
    writer = csv.writer(response)
    writer.writerow(["Sucursal", branch_panel.get("selected_branch_label") or ""])
    writer.writerow(["Periodo pricing", branch_panel.get("latest_period_label") or ""])
    writer.writerow(["Corte YTD", contribution_panel.get("ytd_cutoff_label") or ""])
    writer.writerow([])
    writer.writerow(["Producto", "Familia", "Categoria", "Venta", "Costo", "Gasto comercial", "Contribucion", "Margen %", "Accion", "Recomendacion"])
    for row in branch_panel.get("rows") or []:
        writer.writerow(
            [
                row.get("label"),
                row.get("familia") or "",
                row.get("categoria") or "",
                row.get("sales_total") or 0,
                row.get("cost_total") or 0,
                row.get("commercial_total") or 0,
                row.get("contribution_total") or 0,
                row.get("contribution_pct") or "",
                row.get("bucket") or "",
                row.get("recommendation") or "",
            ]
        )
    return response


def _export_branch_bi_xlsx(branch_panel: dict[str, object], contribution_panel: dict[str, object]) -> HttpResponse:
    branch_slug = str(branch_panel.get("selected_branch_code") or "sucursal").lower()
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Resumen sucursal"
    summary_ws.append(["Sucursal", branch_panel.get("selected_branch_label") or ""])
    summary_ws.append(["Periodo pricing", branch_panel.get("latest_period_label") or ""])
    summary_ws.append(["Corte YTD", contribution_panel.get("ytd_cutoff_label") or ""])
    summary_ws.append([])
    summary_ws.append(["Top sucursales YTD"])
    summary_ws.append(["Sucursal", "Venta", "Contribucion", "No receta"])
    for row in contribution_panel.get("top_rows") or []:
        summary_ws.append(
            [
                row.get("branch_label"),
                float(row.get("sales_total") or 0),
                float(row.get("contribution_total") or 0),
                float(row.get("non_recipe_total") or 0),
            ]
        )

    detail_ws = wb.create_sheet("Pricing sucursal")
    detail_ws.append(["Producto", "Familia", "Categoria", "Venta", "Costo", "Gasto comercial", "Contribucion", "Margen %", "Accion", "Recomendacion"])
    for row in branch_panel.get("rows") or []:
        detail_ws.append(
            [
                row.get("label"),
                row.get("familia") or "",
                row.get("categoria") or "",
                float(row.get("sales_total") or 0),
                float(row.get("cost_total") or 0),
                float(row.get("commercial_total") or 0),
                float(row.get("contribution_total") or 0),
                float(row.get("contribution_pct") or 0),
                row.get("bucket") or "",
                row.get("recommendation") or "",
            ]
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="bi_sucursal_{branch_slug}.xlsx"'
    return response


def _export_branch_bi_pdf(branch_panel: dict[str, object], contribution_panel: dict[str, object]) -> HttpResponse:
    branch_slug = str(branch_panel.get("selected_branch_code") or "sucursal").lower()
    lines = [
        f"Periodo pricing: {branch_panel.get('latest_period_label') or 'Sin snapshot'}",
        f"Corte YTD: {contribution_panel.get('ytd_cutoff_label') or 'Sin corte'}",
        "",
        "Top acciones de pricing",
    ]
    for row in (branch_panel.get("top_actions") or [])[:10]:
        lines.append(
            f"{row.get('bucket')}: {row.get('label')} | Venta ${float(row.get('sales_total') or 0):,.2f} | "
            f"Contribucion ${float(row.get('contribution_total') or 0):,.2f}"
        )
    lines.append("")
    lines.append("Top sucursales YTD")
    for row in (contribution_panel.get("top_rows") or [])[:8]:
        lines.append(
            f"{row.get('branch_label')} | Venta ${float(row.get('sales_total') or 0):,.2f} | "
            f"Contribucion ${float(row.get('contribution_total') or 0):,.2f} | "
            f"No receta ${float(row.get('non_recipe_total') or 0):,.2f}"
        )
    pdf_bytes = _simple_pdf_bytes(
        title=f"BI sucursal: {branch_panel.get('selected_branch_label') or 'Sin sucursal'}",
        lines=lines,
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="bi_sucursal_{branch_slug}.pdf"'
    return response


def _export_branches_bi_csv(contribution_panel: dict[str, object]) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="bi_sucursales_2026.csv"'
    writer = csv.writer(response)
    writer.writerow(["Corte YTD", contribution_panel.get("ytd_cutoff_label") or ""])
    writer.writerow(["Detalle mensual", contribution_panel.get("latest_period_label") or ""])
    writer.writerow([])
    writer.writerow(
        [
            "Sucursal",
            "Codigo",
            "Venta YTD",
            "Contribucion YTD",
            "Margen YTD %",
            "Venta ultimo mes",
            "Contribucion ultimo mes",
            "Margen ultimo mes %",
            "Fabricado %",
            "Reventa %",
            "No receta",
            "No receta reventa",
            "No receta accesorio",
            "No receta servicio",
            "Semaforo",
            "Venta ultimo mes vs prom. YTD %",
            "Contribucion ultimo mes vs prom. YTD %",
            "Brecha margen pp",
        ]
    )
    for row in contribution_panel.get("rows") or []:
        writer.writerow(
            [
                row.get("branch_label"),
                row.get("branch_code"),
                row.get("sales_total") or 0,
                row.get("contribution_total") or 0,
                row.get("contribution_pct") or "",
                row.get("latest_month_sales_total") or 0,
                row.get("latest_month_contribution_total") or 0,
                row.get("latest_month_contribution_pct") or "",
                row.get("fabricated_mix_pct") or "",
                row.get("resale_mix_pct") or "",
                row.get("non_recipe_total") or 0,
                row.get("non_recipe_resale_total") or 0,
                row.get("non_recipe_accessory_total") or 0,
                row.get("non_recipe_service_total") or 0,
                row.get("rank_tone") or "",
                row.get("latest_month_sales_vs_ytd_avg_pct") or "",
                row.get("latest_month_contribution_vs_ytd_avg_pct") or "",
                row.get("latest_month_margin_delta_pp") or "",
            ]
        )
    return response


def _export_branches_bi_xlsx(contribution_panel: dict[str, object]) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sucursales"
    ws.append(["Corte YTD", contribution_panel.get("ytd_cutoff_label") or ""])
    ws.append(["Detalle mensual", contribution_panel.get("latest_period_label") or ""])
    ws.append([])
    ws.append(
        [
            "Sucursal",
            "Codigo",
            "Venta YTD",
            "Contribucion YTD",
            "Margen YTD %",
            "Venta ultimo mes",
            "Contribucion ultimo mes",
            "Margen ultimo mes %",
            "Fabricado %",
            "Reventa %",
            "No receta",
            "No receta reventa",
            "No receta accesorio",
            "No receta servicio",
            "Semaforo",
            "Venta ultimo mes vs prom. YTD %",
            "Contribucion ultimo mes vs prom. YTD %",
            "Brecha margen pp",
        ]
    )
    for row in contribution_panel.get("rows") or []:
        ws.append(
            [
                row.get("branch_label"),
                row.get("branch_code"),
                float(row.get("sales_total") or 0),
                float(row.get("contribution_total") or 0),
                float(row.get("contribution_pct") or 0),
                float(row.get("latest_month_sales_total") or 0),
                float(row.get("latest_month_contribution_total") or 0),
                float(row.get("latest_month_contribution_pct") or 0),
                float(row.get("fabricated_mix_pct") or 0),
                float(row.get("resale_mix_pct") or 0),
                float(row.get("non_recipe_total") or 0),
                float(row.get("non_recipe_resale_total") or 0),
                float(row.get("non_recipe_accessory_total") or 0),
                float(row.get("non_recipe_service_total") or 0),
                row.get("rank_tone") or "",
                float(row.get("latest_month_sales_vs_ytd_avg_pct") or 0),
                float(row.get("latest_month_contribution_vs_ytd_avg_pct") or 0),
                float(row.get("latest_month_margin_delta_pp") or 0),
            ]
        )
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="bi_sucursales_2026.xlsx"'
    return response


def _export_branches_bi_pdf(contribution_panel: dict[str, object]) -> HttpResponse:
    lines = [
        f"Corte YTD: {contribution_panel.get('ytd_cutoff_label') or 'Sin corte'}",
        f"Detalle mensual: {contribution_panel.get('latest_period_label') or 'Sin snapshot'}",
        "",
        "Resumen por sucursal",
    ]
    for row in (contribution_panel.get("rows") or [])[:20]:
        lines.append(
            f"{row.get('branch_label')} | Venta YTD ${float(row.get('sales_total') or 0):,.2f} | "
            f"Contribucion YTD ${float(row.get('contribution_total') or 0):,.2f} | "
            f"Ultimo mes ${float(row.get('latest_month_sales_total') or 0):,.2f} / ${float(row.get('latest_month_contribution_total') or 0):,.2f} | "
            f"No receta ${float(row.get('non_recipe_total') or 0):,.2f}"
        )
    pdf_bytes = _simple_pdf_bytes(title="BI sucursales 2026", lines=lines)
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="bi_sucursales_2026.pdf"'
    return response


def _default_product_closure_month() -> date:
    latest_month = ProductoMonthClosure.objects.order_by("-month_start").values_list("month_start", flat=True).first()
    if latest_month:
        return latest_month
    today = timezone.localdate()
    current_month_start = date(today.year, today.month, 1)
    previous_month_end = current_month_start - timedelta(days=1)
    return date(previous_month_end.year, previous_month_end.month, 1)


def _normalize_product_closure_month(raw_value: str | None) -> date:
    normalized = (raw_value or "").strip()
    if normalized:
        try:
            return date.fromisoformat(f"{normalized}-01")
        except ValueError:
            pass
    return _default_product_closure_month()


def _product_closure_month_options(selected_month_start: date, months_back: int = 10) -> list[dict[str, str]]:
    month_candidates: list[date] = [selected_month_start]
    latest_month = ProductoMonthClosure.objects.order_by("-month_start").values_list("month_start", flat=True).first()
    if latest_month:
        month_candidates.append(latest_month)

    anchor = timezone.localdate().replace(day=1)
    for offset in range(months_back):
        candidate_month_end = anchor - timedelta(days=offset * 31)
        candidate_month_start = date(candidate_month_end.year, candidate_month_end.month, 1)
        month_candidates.append(candidate_month_start)

    unique_months: list[date] = []
    seen: set[str] = set()
    for month_start in sorted(month_candidates, reverse=True):
        token = month_start.strftime("%Y-%m")
        if token in seen:
            continue
        seen.add(token)
        unique_months.append(month_start)

    return [
        {
            "value": month_start.strftime("%Y-%m"),
            "label": month_start.strftime("%B %Y").capitalize(),
        }
        for month_start in unique_months
    ]


def _product_closure_status_tone(status: str) -> str:
    if status == ProductoMonthClosure.STATUS_LOCKED:
        return "success"
    if status == ProductoMonthClosure.STATUS_BUILT:
        return "warning"
    return "neutral"


def _build_product_closure_context(selected_month_start: date) -> dict[str, object]:
    closure = (
        ProductoMonthClosure.objects.select_related("built_by").prefetch_related("lines", "lines__receta_padre")
        .filter(month_start=selected_month_start)
        .order_by("-id")
        .first()
    )
    lines = list(closure.lines.select_related("receta_padre").all()) if closure else []

    total_opening = sum((Decimal(str(line.inventario_inicial_teorico or 0)) for line in lines), Decimal("0"))
    total_production = sum((Decimal(str(line.produccion_mes or 0)) for line in lines), Decimal("0"))
    total_sales = sum((Decimal(str(line.venta_total_equivalente or 0)) for line in lines), Decimal("0"))
    total_waste = sum((Decimal(str(line.merma_total_equivalente or 0)) for line in lines), Decimal("0"))
    total_ending = sum((Decimal(str(line.inventario_final_teorico or 0)) for line in lines), Decimal("0"))
    total_direct_sales = sum((Decimal(str(line.venta_directa_enteros or 0)) for line in lines), Decimal("0"))
    total_derived_sales = sum((Decimal(str(line.venta_derivada_equivalente or 0)) for line in lines), Decimal("0"))
    total_direct_waste = sum((Decimal(str(line.merma_directa_enteros or 0)) for line in lines), Decimal("0"))
    total_derived_waste = sum((Decimal(str(line.merma_derivada_equivalente or 0)) for line in lines), Decimal("0"))

    conversion_rows = [
        line
        for line in lines
        if Decimal(str(line.venta_derivada_equivalente or 0)) > 0
        or Decimal(str(line.merma_derivada_equivalente or 0)) > 0
    ]
    conversion_rows.sort(
        key=lambda line: (
            Decimal(str(line.venta_derivada_equivalente or 0)) + Decimal(str(line.merma_derivada_equivalente or 0)),
            line.receta_padre.nombre.lower(),
        ),
        reverse=True,
    )

    catalog_issue_rows = [line for line in lines if line.has_catalog_issue]
    recent_closures = (
        ProductoMonthClosure.objects.prefetch_related("lines")
        .order_by("-month_start", "-id")[:6]
    )

    recent_rows: list[dict[str, object]] = []
    for recent in recent_closures:
        recent_lines = list(recent.lines.all())
        recent_rows.append(
            {
                "month_label": recent.month_start.strftime("%b %Y").capitalize(),
                "month_value": recent.month_start.strftime("%Y-%m"),
                "status": recent.get_status_display(),
                "tone": _product_closure_status_tone(recent.status),
                "line_count": len(recent_lines),
                "ending_inventory": sum(
                    (Decimal(str(line.inventario_final_teorico or 0)) for line in recent_lines),
                    Decimal("0"),
                ),
                "sales_total": sum(
                    (Decimal(str(line.venta_total_equivalente or 0)) for line in recent_lines),
                    Decimal("0"),
                ),
            }
        )

    notes_rows = [row.strip() for row in (closure.notes or "").splitlines() if row.strip()] if closure else []
    opening_meta = (closure.metadata or {}).get("opening_meta", {}) if closure else {}
    validation = (closure.metadata or {}).get("validation", {}) if closure else {}
    unmatched_products = list(opening_meta.get("unmatched_products") or [])[:8]
    lock_event = (closure.metadata or {}).get("lock_event", {}) if closure else {}
    lock_guard_errors: list[str] = []
    if closure:
        if closure.is_locked:
            lock_guard_errors.append("El cierre ya fue bloqueado.")
        elif closure.status != ProductoMonthClosure.STATUS_BUILT:
            lock_guard_errors.append("El cierre debe estar construido antes de bloquearse.")
        elif not lines:
            lock_guard_errors.append("El cierre no tiene lineas para bloquear.")
        elif catalog_issue_rows:
            lock_guard_errors.append("Existen incidencias de catalogo pendientes en las lineas del cierre.")
        if unmatched_products:
            lock_guard_errors.append("Existen productos del opening sin homologacion Point -> ERP.")

    exception_rows: list[dict[str, str]] = []
    if validation.get("snapshot_fallback_used"):
        effective_date = opening_meta.get("snapshot_effective_date") or "sin fecha"
        exception_rows.append(
            {
                "tone": "warning",
                "title": "Snapshot fallback",
                "detail": f"El opening uso snapshot previo dentro de tolerancia con fecha efectiva {effective_date}.",
            }
        )
    if unmatched_products:
        exception_rows.append(
            {
                "tone": "danger",
                "title": "Productos sin homologacion",
                "detail": f"El opening trae {len(unmatched_products)} producto(s) Point sin homologar a receta ERP.",
            }
        )
    if catalog_issue_rows:
        exception_rows.append(
            {
                "tone": "danger",
                "title": "Incidencias de catalogo en lineas",
                "detail": f"El cierre tiene {len(catalog_issue_rows)} linea(s) con derivadas o catalogo pendientes.",
            }
        )
    if closure and closure.is_locked:
        exception_rows.append(
            {
                "tone": "success",
                "title": "Mes protegido",
                "detail": "El cierre ya fue bloqueado y queda como base canónica para el siguiente periodo.",
            }
        )

    return {
        "selected_month": selected_month_start.strftime("%Y-%m"),
        "selected_month_label": selected_month_start.strftime("%B %Y").capitalize(),
        "month_options": _product_closure_month_options(selected_month_start),
        "closure": closure,
        "closure_lines": lines,
        "closure_status_tone": _product_closure_status_tone(closure.status) if closure else "warning",
        "total_opening": total_opening,
        "total_production": total_production,
        "total_sales": total_sales,
        "total_waste": total_waste,
        "total_ending": total_ending,
        "total_direct_sales": total_direct_sales,
        "total_derived_sales": total_derived_sales,
        "total_direct_waste": total_direct_waste,
        "total_derived_waste": total_derived_waste,
        "catalog_issue_count": len(catalog_issue_rows),
        "conversion_rows": conversion_rows[:8],
        "catalog_issue_rows": catalog_issue_rows[:8],
        "recent_closure_rows": recent_rows,
        "notes_rows": notes_rows,
        "unmatched_products": unmatched_products,
        "lock_event": lock_event,
        "lock_guard_errors": lock_guard_errors,
        "validation": validation,
        "exception_rows": exception_rows,
    }


def _export_product_closure_csv(context: dict[str, object]) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="cierre_producto_{context["selected_month"]}.csv"'
    writer = csv.writer(response)
    writer.writerow(["Mes", context["selected_month"]])
    writer.writerow(["Estado", context["closure"].status if context.get("closure") else ""])
    writer.writerow(["Opening source", context["closure"].opening_source if context.get("closure") else ""])
    writer.writerow(["Inventario inicial", context["total_opening"]])
    writer.writerow(["Produccion", context["total_production"]])
    writer.writerow(["Venta equivalente", context["total_sales"]])
    writer.writerow(["Merma equivalente", context["total_waste"]])
    writer.writerow(["Inventario final", context["total_ending"]])
    writer.writerow([])
    writer.writerow(
        [
            "Receta padre",
            "Codigo point",
            "Inicial",
            "Produccion",
            "Venta directa",
            "Venta derivada",
            "Merma total",
            "Final teorico",
            "Catalog issue",
            "Catalog issue note",
        ]
    )
    for line in context.get("closure_lines") or []:
        writer.writerow(
            [
                line.receta_padre.nombre,
                line.receta_padre.codigo_point,
                line.inventario_inicial_teorico,
                line.produccion_mes,
                line.venta_directa_enteros,
                line.venta_derivada_equivalente,
                line.merma_total_equivalente,
                line.inventario_final_teorico,
                "SI" if line.has_catalog_issue else "NO",
                line.catalog_issue_note,
            ]
        )
    return response


def _export_product_closure_xlsx(context: dict[str, object]) -> HttpResponse:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Resumen"
    summary_ws.append(["Mes", context["selected_month"]])
    summary_ws.append(["Estado", context["closure"].status if context.get("closure") else ""])
    summary_ws.append(["Opening source", context["closure"].opening_source if context.get("closure") else ""])
    summary_ws.append(["Inventario inicial", float(context["total_opening"])])
    summary_ws.append(["Produccion", float(context["total_production"])])
    summary_ws.append(["Venta equivalente", float(context["total_sales"])])
    summary_ws.append(["Merma equivalente", float(context["total_waste"])])
    summary_ws.append(["Inventario final", float(context["total_ending"])])

    detail_ws = wb.create_sheet("Detalle")
    detail_ws.append(
        [
            "Receta padre",
            "Codigo point",
            "Inicial",
            "Produccion",
            "Venta directa",
            "Venta derivada",
            "Merma total",
            "Final teorico",
            "Catalog issue",
            "Catalog issue note",
        ]
    )
    for line in context.get("closure_lines") or []:
        detail_ws.append(
            [
                line.receta_padre.nombre,
                line.receta_padre.codigo_point,
                float(line.inventario_inicial_teorico or 0),
                float(line.produccion_mes or 0),
                float(line.venta_directa_enteros or 0),
                float(line.venta_derivada_equivalente or 0),
                float(line.merma_total_equivalente or 0),
                float(line.inventario_final_teorico or 0),
                "SI" if line.has_catalog_issue else "NO",
                line.catalog_issue_note,
            ]
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="cierre_producto_{context["selected_month"]}.xlsx"'
    return response


@login_required
def cierre_producto(request: HttpRequest) -> HttpResponse:
    if not can_view_product_closure(request.user):
        raise PermissionDenied("No tienes permisos para ver Reportes.")

    selected_month_start = _normalize_product_closure_month(
        request.POST.get("month") if request.method == "POST" else request.GET.get("month")
    )

    if request.method == "POST":
        action = (request.POST.get("action") or "build").strip().lower()
        existing_closure = ProductoMonthClosure.objects.filter(month_start=selected_month_start).order_by("-id").first()
        service = ProductMonthClosureService()
        if action == "lock":
            if existing_closure is None:
                messages.error(request, f"No existe un cierre construido para {selected_month_start:%Y-%m}.")
            elif not can_lock_product_closure(request.user):
                messages.error(
                    request,
                    "Solo Dirección General o Administración pueden bloquear cierres mensuales.",
                )
            else:
                try:
                    service.lock(
                        closure=existing_closure,
                        locked_by=request.user,
                        reason="ui_lock",
                        note=request.POST.get("approval_note") or "",
                        channel="ui",
                    )
                    messages.success(
                        request,
                        f"El cierre {selected_month_start:%Y-%m} quedó bloqueado para proteger la conciliación mensual.",
                    )
                except ProductMonthClosureError as exc:
                    messages.error(request, str(exc))
        else:
            if existing_closure is not None:
                if not can_rebuild_product_closure(request.user):
                    messages.warning(
                        request,
                        f"El cierre {selected_month_start:%Y-%m} ya existe. La vista no permite rebuild manual para proteger la conciliación.",
                    )
                else:
                    try:
                        service.build(
                            month=selected_month_start,
                            rebuild=True,
                            built_by=request.user,
                            approval_channel="ui",
                        )
                        messages.success(
                            request,
                            f"Se reconstruyó el cierre teórico de producto Point para {selected_month_start:%Y-%m}.",
                        )
                    except ProductMonthClosureError as exc:
                        messages.error(request, str(exc))
            else:
                if not can_build_product_closure(request.user):
                    messages.error(
                        request,
                        "No tienes permisos para construir cierres mensuales de producto.",
                    )
                    return redirect(f"{reverse('reportes:cierre_producto')}?month={selected_month_start:%Y-%m}")
                try:
                    service.build(
                        month=selected_month_start,
                        built_by=request.user,
                        approval_channel="ui",
                    )
                    messages.success(
                        request,
                        f"Se construyó el cierre teórico de producto Point para {selected_month_start:%Y-%m}.",
                    )
                except ProductMonthClosureError as exc:
                    messages.error(request, str(exc))
        return redirect(f"{reverse('reportes:cierre_producto')}?month={selected_month_start:%Y-%m}")

    context = _build_product_closure_context(selected_month_start)
    export_format = (request.GET.get("export") or "").lower()
    if context.get("closure") and export_format == "csv":
        return _export_product_closure_csv(context)
    if context.get("closure") and export_format == "xlsx":
        return _export_product_closure_xlsx(context)
    context["can_lock_closure"] = can_lock_product_closure(request.user)
    context["can_build_closure"] = can_build_product_closure(request.user)
    context["can_rebuild_closure"] = can_rebuild_product_closure(request.user)
    return render(request, "reportes/cierre_producto_point.html", context)


@login_required
def bi(request: HttpRequest) -> HttpResponse:
    if not can_view_reportes(request.user):
        raise PermissionDenied("No tienes permisos para ver Reportes.")

    try:
        period_days = int(request.GET.get("period_days") or "90")
    except (TypeError, ValueError):
        period_days = 90
    try:
        months_window = int(request.GET.get("months") or "6")
    except (TypeError, ValueError):
        months_window = 6
    try:
        branch_id = int(request.GET.get("branch_id") or "0") or None
    except (TypeError, ValueError):
        branch_id = None
    try:
        budget_month = int(request.GET.get("budget_month") or "0") or None
    except (TypeError, ValueError):
        budget_month = None
    action_filter = (request.GET.get("action") or "").strip()
    bi_runtime_cache: dict[str, object] = {}
    snapshot = _bi_cached_value(
        runtime_cache=bi_runtime_cache,
        section="snapshot",
        builder=lambda: compute_bi_snapshot(period_days=period_days, months_window=months_window),
        parts=(period_days, months_window),
    )
    executive_panels = build_executive_bi_panels(
        months=months_window,
        branch_id=branch_id,
        action_filter=action_filter,
        budget_month=budget_month,
    )

    export_format = (request.GET.get("export") or "").lower()
    if branch_id and export_format == "csv":
        return _export_branch_bi_csv(executive_panels["branch_pricing_panel"], executive_panels["branch_contribution_panel"])
    if branch_id and export_format == "xlsx":
        return _export_branch_bi_xlsx(executive_panels["branch_pricing_panel"], executive_panels["branch_contribution_panel"])
    if branch_id and export_format == "pdf":
        return _export_branch_bi_pdf(executive_panels["branch_pricing_panel"], executive_panels["branch_contribution_panel"])
    if export_format == "branches_csv":
        return _export_branches_bi_csv(executive_panels["branch_contribution_panel"])
    if export_format == "branches_xlsx":
        return _export_branches_bi_xlsx(executive_panels["branch_contribution_panel"])
    if export_format == "branches_pdf":
        return _export_branches_bi_pdf(executive_panels["branch_contribution_panel"])
    if export_format == "csv":
        return _export_bi_csv(snapshot)
    if export_format == "xlsx":
        return _export_bi_xlsx(snapshot)

    context = {
        "snapshot": snapshot,
        "executive_panels": executive_panels,
        "forecast_panel": executive_panels["forecast_panel"],
        "yoy_panel": executive_panels["yoy_panel"],
        "profitability_panel": executive_panels["profitability_panel"],
        "branch_contribution_panel": executive_panels["branch_contribution_panel"],
        "branch_pricing_panel": executive_panels["branch_pricing_panel"],
        "budget_operating_panel": executive_panels["budget_operating_panel"],
        "production_sales_panel": executive_panels["production_sales_panel"],
        "central_flow_panel": executive_panels["central_flow_panel"],
        "inventory_ledger_panel": executive_panels["inventory_ledger_panel"],
        "daily_sales_snapshot": _bi_cached_value(
            runtime_cache=bi_runtime_cache,
            section="daily-sales-snapshot",
            builder=_bi_daily_sales_snapshot,
            parts=(timezone.localdate().isoformat(),),
        ),
        "branch_weekday_rows": _bi_cached_value(
            runtime_cache=bi_runtime_cache,
            section="branch-weekday-rows",
            builder=_bi_branch_weekday_comparisons,
            parts=(timezone.localdate().isoformat(),),
        ),
        "product_weekday_rows": _bi_cached_value(
            runtime_cache=bi_runtime_cache,
            section="product-weekday-rows",
            builder=_bi_product_weekday_comparisons,
            parts=(timezone.localdate().isoformat(),),
        ),
        "purchase_snapshot": _bi_cached_value(
            runtime_cache=bi_runtime_cache,
            section="purchase-snapshot",
            builder=_bi_purchase_snapshot,
            parts=(timezone.localdate().isoformat(),),
        ),
        "inventory_snapshot": _bi_cached_value(
            runtime_cache=bi_runtime_cache,
            section="inventory-snapshot",
            builder=_bi_inventory_snapshot,
            parts=(timezone.localdate().isoformat(),),
        ),
        "production_snapshot": _bi_cached_value(
            runtime_cache=bi_runtime_cache,
            section="production-snapshot",
            builder=_bi_production_snapshot,
            parts=(timezone.localdate().isoformat(),),
        ),
        "production_summary": _bi_cached_value(
            runtime_cache=bi_runtime_cache,
            section="production-summary",
            builder=lambda: _bi_production_summary(snapshot["range"]["from"], snapshot["range"]["to"]),
            parts=(snapshot["range"]["from"].isoformat(), snapshot["range"]["to"].isoformat()),
        ),
        "waste_summary": _bi_cached_value(
            runtime_cache=bi_runtime_cache,
            section="waste-summary",
            builder=lambda: _bi_waste_summary(snapshot["range"]["from"], snapshot["range"]["to"]),
            parts=(snapshot["range"]["from"].isoformat(), snapshot["range"]["to"].isoformat()),
        ),
        "forecast_summary": _bi_cached_value(
            runtime_cache=bi_runtime_cache,
            section="forecast-summary",
            builder=lambda: _bi_forecast_summary(snapshot["range"]["to"].strftime("%Y-%m")),
            parts=(snapshot["range"]["to"].strftime("%Y-%m"),),
        ),
        "supply_watchlist": _bi_cached_value(
            runtime_cache=bi_runtime_cache,
            section="supply-watchlist",
            builder=_bi_supply_watchlist,
            parts=(timezone.localdate().isoformat(),),
        ),
        "ventas_historicas_summary": _bi_cached_value(
            runtime_cache=bi_runtime_cache,
            section="ventas-historicas-summary",
            builder=_ventas_historicas_bi_summary,
            parts=(timezone.localdate().isoformat(),),
        ),
        "bi_force_refresh_reference_date": timezone.localdate().isoformat(),
        "bi_force_refresh_lookback_days": 7,
        "current_year": executive_panels["yoy_panel"]["current_year"],
        "prev_year": executive_panels["yoy_panel"]["prev_year"],
        "prev2_year": executive_panels["yoy_panel"]["prev2_year"],
        "period_days": snapshot["range"]["days"],
        "months_window": snapshot["range"]["months_window"],
        "selected_branch_id": branch_id,
        "selected_budget_month": budget_month,
        "selected_action": action_filter,
        "enterprise_chain": _reportes_enterprise_chain(
            focus="bi",
            open_count=len(snapshot.get("series_mensual", [])),
            blocked_count=0,
        ),
        "document_stage_rows": _reportes_document_stage_rows(
            focus="bi",
            open_count=len(snapshot.get("series_mensual", [])),
            blocked_count=0,
            total_count=len(snapshot.get("series_mensual", [])),
        ),
        "erp_governance_rows": _reportes_governance_rows(
            _reportes_document_stage_rows(
                focus="bi",
                open_count=len(snapshot.get("series_mensual", [])),
                blocked_count=0,
                total_count=len(snapshot.get("series_mensual", [])),
            )
        ),
        "operational_health_cards": _reportes_operational_health_cards(
            focus="bi",
            open_count=len(snapshot.get("series_mensual", [])),
            blocked_count=0,
            total_count=len(snapshot.get("series_mensual", [])),
        ),
        "maturity_summary": _reportes_maturity_summary(
            chain=_reportes_enterprise_chain(
                focus="bi",
                open_count=len(snapshot.get("series_mensual", [])),
                blocked_count=0,
            ),
            default_url=reverse("reportes:bi"),
        ),
        "handoff_map": _reportes_handoff_map(
            focus="bi",
            blocked_count=0,
            open_count=len(snapshot.get("series_mensual", [])),
            total_count=len(snapshot.get("series_mensual", [])),
            default_url=reverse("reportes:bi"),
        ),
        "release_gate_rows": _reportes_release_gate_rows(
            focus="bi",
            open_count=len(snapshot.get("series_mensual", [])),
            blocked_count=0,
            total_count=len(snapshot.get("series_mensual", [])),
        ),
    }
    context["sales_branch_bar_rows"] = _bi_bar_rows(
        list(context["daily_sales_snapshot"].get("top_branches") or []),
        label_key="label",
        secondary_key="secondary",
        value_key="amount",
    )
    context["sales_product_bar_rows"] = _bi_bar_rows(
        list(context["daily_sales_snapshot"].get("top_products") or []),
        label_key="label",
        value_key="amount",
    )
    context["monthly_sales_rows"] = _bi_monthly_sales_rows(snapshot)
    context["monthly_margin_rows"] = _bi_monthly_margin_rows(snapshot)
    context["supplier_bar_rows"] = _bi_bar_rows(
        list(snapshot.get("top_proveedores") or []),
        label_key="proveedor__nombre",
        value_key="total",
    )
    context["consumption_bar_rows"] = _bi_bar_rows(
        list(snapshot.get("top_insumos_consumo") or []),
        label_key="insumo__nombre",
        value_key="total",
    )
    context["production_product_bar_rows"] = _bi_bar_rows(
        list(context["production_summary"].get("top_products") or []),
        label_key="label",
        value_key="value",
    )
    context["waste_branch_bar_rows"] = _bi_bar_rows(
        list(context["waste_summary"].get("branch_rows") or []),
        label_key="label",
        secondary_key="secondary",
        value_key="value",
    )
    context["waste_cedis_bar_rows"] = _bi_bar_rows(
        list(context["waste_summary"].get("cedis_rows") or []),
        label_key="label",
        secondary_key="secondary",
        value_key="value",
    )
    context["forecast_gap_bar_rows"] = _bi_bar_rows(
        list(context["forecast_summary"].get("top_rows") or []),
        label_key="label",
        secondary_key="secondary",
        value_key="value",
    )
    context["branch_weekday_bar_rows"] = _bi_comparison_bar_rows(
        list(context["branch_weekday_rows"] or []),
        label_key="branch_code",
        secondary_key="branch_name",
    )
    context["product_weekday_bar_rows"] = _bi_comparison_bar_rows(
        list(context["product_weekday_rows"] or []),
        label_key="recipe_name",
    )
    context["daily_decision_rows"] = _bi_daily_decisions(
        daily_sales_snapshot=context["daily_sales_snapshot"],
        branch_weekday_rows=context["branch_weekday_rows"],
        product_weekday_rows=context["product_weekday_rows"],
        purchase_snapshot=context["purchase_snapshot"],
        inventory_snapshot=context["inventory_snapshot"],
        production_snapshot=context["production_snapshot"],
        waste_summary=context["waste_summary"],
        forecast_summary=context["forecast_summary"],
        supply_watchlist=context["supply_watchlist"],
    )
    context["release_gate_completion"] = _reportes_release_gate_completion(context["release_gate_rows"])
    context["critical_path_rows"] = _reportes_critical_path_rows(context["enterprise_chain"])
    context["executive_radar_rows"] = _reportes_executive_radar_rows(
        context["erp_governance_rows"],
        default_owner="Dirección General",
        fallback_url=reverse("reportes:bi"),
    )
    context["erp_command_center"] = _reportes_command_center(
        governance_rows=context["erp_governance_rows"],
        maturity_summary=context["maturity_summary"],
        default_owner="Dirección General",
    )
    return render(request, "reportes/bi.html", context)


@login_required
def production_orders(request: HttpRequest) -> HttpResponse:
    if not can_view_reportes(request.user):
        raise PermissionDenied

    target_date = _parse_ui_date(request.GET.get("fecha") or request.POST.get("fecha"))
    if request.method == "POST":
        if not can_manage_orquestacion(request.user):
            raise PermissionDenied
        action = (request.POST.get("action") or "").strip().lower()
        if action == "generate":
            result = generate_daily_production_orders(target_date, created_by=request.user)
            messages.success(
                request,
                f"Órdenes generadas: nuevas {result['generated_orders']}, actualizadas {result['updated_orders']}, líneas {result['lines']}.",
            )
        elif action == "generate_purchases":
            order = ProductionOrder.objects.filter(pk=request.POST.get("order_id"), fecha=target_date).first()
            branch_id = order.sucursal_id if order else None
            result = generate_purchase_requests_from_production(
                target_date,
                sucursal_id=branch_id,
                actor=request.user,
            )
            messages.success(
                request,
                (
                    f"Solicitudes de compra generadas: nuevas {result['generated']}, "
                    f"actualizadas {result['updated']}, líneas {result['lines']}."
                ),
            )
        elif action == "generate_alerts":
            result = generate_operational_alerts(target_date=target_date)
            messages.success(
                request,
                f"Alertas actualizadas: {result['created_or_updated']} registro(s), críticas {result['critical']}.",
            )
        elif action == "resolve_alert":
            alert = Alert.objects.filter(pk=request.POST.get("alert_id"), fecha=target_date).first()
            if alert is None:
                messages.error(request, "No se encontró la alerta a resolver.")
            else:
                resolve_alert(
                    alert=alert,
                    resolved_by=request.user,
                    resolution_note=request.POST.get("resolution_note") or "",
                    impacto_real=_to_decimal(request.POST.get("impacto_real")),
                )
                messages.success(request, f"Alerta {alert.id} resuelta.")
        elif action == "refresh_metrics":
            result = rebuild_operations_metrics(target_date=target_date)
            messages.success(
                request,
                (
                    f"Métricas DG actualizadas. Adopción {result['adoption_pct']}%, "
                    f"merma {result['merma_total']}."
                ),
            )
        else:
            order = ProductionOrder.objects.filter(pk=request.POST.get("order_id"), fecha=target_date).first()
            if order is None:
                messages.error(request, "No se encontró la orden solicitada para la fecha operativa.")
                return redirect(f"{reverse('reportes:production_orders')}?fecha={target_date.isoformat()}")
            if action == "approve":
                approved_quantities = {
                    line.receta_id: _to_decimal(request.POST.get(f"approved_{line.id}"))
                    for line in order.lines.all()
                    if request.POST.get(f"approved_{line.id}") not in (None, "")
                }
                approve_production_order(order, approved_by=request.user, approved_quantities=approved_quantities)
                messages.success(request, f"Orden {order.id} aprobada.")
            elif action == "release":
                release_production_order(order)
                messages.success(request, f"Orden {order.id} liberada a producción.")
            elif action == "execute":
                executed_quantities = {
                    line.receta_id: _to_decimal(request.POST.get(f"executed_{line.id}"))
                    for line in order.lines.all()
                    if request.POST.get(f"executed_{line.id}") not in (None, "")
                }
                execute_production_order(order, executed_quantities=executed_quantities)
                sync_result = sync_production_execution_logs(
                    target_date=target_date,
                    sucursal_id=order.sucursal_id,
                    actor=request.user,
                )
                messages.success(
                    request,
                    f"Orden {order.id} ejecutada y bitácora sincronizada ({sync_result['logs']} registro(s)).",
                )
            elif action == "sync":
                sync_result = sync_production_execution_logs(
                    target_date=target_date,
                    sucursal_id=order.sucursal_id,
                    actor=request.user,
                )
                messages.success(request, f"Bitácora sincronizada ({sync_result['logs']} registro(s)).")
            else:
                messages.error(request, "Acción operativa no soportada.")
        return redirect(f"{reverse('reportes:production_orders')}?fecha={target_date.isoformat()}")

    orders = list(
        ProductionOrder.objects.filter(fecha=target_date)
        .select_related("sucursal", "approved_by", "created_by")
        .prefetch_related("lines__receta")
        .order_by("sucursal__codigo", "id")
    )
    purchase_requests = list(list_auto_purchase_snapshots(target_date=target_date))
    forecast_context = build_daily_forecast_context(target_date=target_date, top_n=24)
    projection_supply_context = build_projection_supply_context(
        target_date=target_date,
        top_n=24,
        forecast_context=forecast_context,
    )
    supply_context = build_production_supply_context(target_date=target_date, orders=orders)
    supply_by_order = supply_context.get("orders") or {}
    for order in orders:
        order.supply_reconciliation = supply_by_order.get(
            int(order.id),
            {
                "item_count": 0,
                "shortage_items": 0,
                "covered_items": 0,
                "generated_purchase_rows": 0,
                "coverage_rows_pct": Decimal("0.00"),
                "items": [],
                "has_shortage": False,
                "inventory_scope": "GLOBAL_INSUMO",
            },
        )
    active_alerts = list(
        Alert.objects.select_related("sucursal", "receta", "insumo")
        .filter(fecha=target_date, resuelta=False)
        .order_by("-impacto_estimado", "severidad", "-id")[:24]
    )
    resolved_alerts = list(
        Alert.objects.select_related("sucursal", "receta", "insumo", "resolved_by")
        .filter(fecha=target_date, resuelta=True)
        .order_by("-resolved_at", "-impacto_estimado", "-id")[:12]
    )
    operations_metrics = OperationsMetricSnapshot.objects.filter(fecha=target_date).first()
    branch_metrics = []
    if operations_metrics and isinstance(operations_metrics.payload, dict):
        branch_metrics = [
            {"branch_code": branch_code, **(values or {})}
            for branch_code, values in sorted((operations_metrics.payload.get("by_branch") or {}).items())
        ]
    summary = {
        "orders": len(orders),
        "lines": sum(len(order.lines.all()) for order in orders),
        "proposed": sum(1 for order in orders if order.status == ProductionOrder.STATUS_PROPOSED),
        "approved": sum(1 for order in orders if order.status == ProductionOrder.STATUS_APPROVED),
        "released": sum(1 for order in orders if order.status == ProductionOrder.STATUS_RELEASED),
        "executed": sum(1 for order in orders if order.status == ProductionOrder.STATUS_EXECUTED),
        "purchase_requests": len(purchase_requests),
        "critical_alerts": sum(1 for alert in active_alerts if alert.severidad == Alert.SEVERITY_HIGH),
    }
    context = {
        "selected_date": target_date,
        "orders": orders,
        "purchase_requests": purchase_requests,
        "projection_supply_context": projection_supply_context,
        "supply_context": supply_context,
        "active_alerts": active_alerts,
        "resolved_alerts": resolved_alerts,
        "operations_metrics": operations_metrics,
        "branch_metrics": branch_metrics,
        "summary": summary,
        "can_manage_orders": can_manage_orquestacion(request.user),
    }
    return render(request, "reportes/production_orders.html", context)
