"""Pruebas de consolidación del real en el presupuesto maestro."""

from datetime import date
from decimal import Decimal
from io import StringIO

from django.core.management import call_command
from django.test import TestCase
from django.utils import timezone

from core.models import Sucursal
from pos_bridge.models import PointBranch
from pos_bridge.models.sales_pipeline import PointSalesDailyProductFact
from reportes.models import (
    AreaPresupuesto,
    CategoriaGasto,
    CentroCosto,
    EmpresaResultadoMensual,
    GastoOperativoMensual,
    LineaPresupuestoMensual,
    ReglaFuenteRubro,
    RubroPresupuesto,
)
from reportes.services_presupuesto_maestro import PresupuestoMaestroService
from reportes.services_presupuesto_real import (
    PresupuestoRealConsolidacionService,
    migrar_fuentes_legadas,
)
from rrhh.models import Empleado, NominaLinea, NominaPeriodo


class PresupuestoRealConsolidacionTests(TestCase):
    """Valida las fuentes automáticas y la protección de capturas manuales."""

    @classmethod
    def setUpTestData(cls):
        cls.periodo = date(2026, 3, 1)
        cls.sucursal = Sucursal.objects.create(codigo="GVE01", nombre="Centro")
        cls.otra_sucursal = Sucursal.objects.create(codigo="GVE02", nombre="Norte")
        cls.area = AreaPresupuesto.objects.create(nombre="Pruebas", codigo="pruebas")
        cls.categoria = CategoriaGasto.objects.create(
            codigo="PRUEBA_REAL",
            nombre="Categoría prueba",
            capa_objetivo=CategoriaGasto.CAPA_EMPRESA,
        )
        cls.otra_categoria = CategoriaGasto.objects.create(
            codigo="OTRA_PRUEBA_REAL",
            nombre="Otra categoría",
            capa_objetivo=CategoriaGasto.CAPA_EMPRESA,
        )
        cls.centro = CentroCosto.objects.create(
            codigo="CC-GVE01",
            nombre="Centro sucursal",
            tipo=CentroCosto.TIPO_SUCURSAL,
            sucursal=cls.sucursal,
        )
        cls.otro_centro = CentroCosto.objects.create(
            codigo="CC-GVE02",
            nombre="Centro otra sucursal",
            tipo=CentroCosto.TIPO_SUCURSAL,
            sucursal=cls.otra_sucursal,
        )
        cls.corporativo = CentroCosto.objects.create(
            codigo="CC-CORP",
            nombre="Corporativo",
            tipo=CentroCosto.TIPO_CORPORATIVO,
        )

    def crear_linea(self, concepto="Rubro prueba", *, sucursal=None, monto_real=None, fuente_real="", area=None, tipo=None):
        rubro = RubroPresupuesto.objects.create(
            area=area or self.area,
            concepto=concepto,
            tipo=tipo or RubroPresupuesto.TIPO_EGRESO,
            sucursal=sucursal,
        )
        linea = LineaPresupuestoMensual.objects.create(
            rubro=rubro,
            periodo=self.periodo,
            monto_presupuesto=Decimal("1000"),
            monto_real=monto_real,
            fuente_real=fuente_real,
        )
        return rubro, linea

    def crear_gasto(self, monto, *, periodo=None, centro=None, categoria=None, tipo_dato=None):
        return GastoOperativoMensual.objects.create(
            periodo=periodo or self.periodo,
            centro_costo=centro or self.centro,
            categoria_gasto=categoria or self.categoria,
            monto=Decimal(str(monto)),
            tipo_dato=tipo_dato or GastoOperativoMensual.TIPO_DATO_REAL,
        )

    def crear_regla_gasto(self, rubro, **kwargs):
        return ReglaFuenteRubro.objects.create(
            rubro=rubro,
            tipo_fuente=ReglaFuenteRubro.FUENTE_GASTO_OPERATIVO,
            categoria_gasto=self.categoria,
            **kwargs,
        )

    def consolidar(self, **kwargs):
        return PresupuestoRealConsolidacionService().consolidar(periodo=self.periodo, **kwargs)

    def test_gasto_operativo_suma_solo_reales_del_periodo_categoria_y_sucursal(self):
        """GASTO_OPERATIVO ignora presupuesto, otros meses, categorías y sucursales."""
        rubro, linea = self.crear_linea(sucursal=self.sucursal)
        self.crear_regla_gasto(rubro)
        self.crear_gasto("100")
        self.crear_gasto("25")
        self.crear_gasto("900", tipo_dato=GastoOperativoMensual.TIPO_DATO_PRESUPUESTO)
        self.crear_gasto("800", periodo=date(2026, 2, 1))
        self.crear_gasto("700", centro=self.otro_centro)
        self.crear_gasto("600", categoria=self.otra_categoria)

        self.consolidar()

        linea.refresh_from_db()
        self.assertEqual(linea.monto_real, Decimal("125.00"))
        self.assertEqual(linea.fuente_real, "AUTO:GASTO_OPERATIVO")

    def test_gasto_operativo_filtra_centro_corporativo_sin_sucursal(self):
        """El filtro centro_tipo limita gastos de rubros sin sucursal."""
        rubro, linea = self.crear_linea(concepto="Corporativo")
        self.crear_regla_gasto(rubro, filtros={"centro_tipo": "CORPORATIVO"})
        self.crear_gasto("310", centro=self.corporativo)
        self.crear_gasto("999", centro=self.centro)

        self.consolidar()

        linea.refresh_from_db()
        self.assertEqual(linea.monto_real, Decimal("310.00"))

    def test_nomina_filtra_campo_estatus_mes_departamento_y_sucursal(self):
        """NOMINA suma salario base cerrado/pagado del departamento y sucursal."""
        rubro, linea = self.crear_linea(concepto="Sueldos ventas", sucursal=self.sucursal)
        ReglaFuenteRubro.objects.create(
            rubro=rubro,
            tipo_fuente=ReglaFuenteRubro.FUENTE_NOMINA,
            filtros={"campo_monto": "salario_base", "departamento": "ventas"},
        )
        empleado_valido = Empleado.objects.create(
            codigo="EMP-REAL-1", nombre="Venta válida", departamento=Empleado.DEP_VENTAS, sucursal_ref=self.sucursal
        )
        empleado_otro_depto = Empleado.objects.create(
            codigo="EMP-REAL-2", nombre="Producción", departamento=Empleado.DEP_PRODUCCION, sucursal_ref=self.sucursal
        )
        empleado_otra_sucursal = Empleado.objects.create(
            codigo="EMP-REAL-3", nombre="Venta norte", departamento=Empleado.DEP_VENTAS, sucursal_ref=self.otra_sucursal
        )

        def agregar_linea(folio, fin, estatus, empleado, monto):
            periodo = NominaPeriodo.objects.create(
                folio=folio, fecha_inicio=fin.replace(day=1), fecha_fin=fin, estatus=estatus
            )
            NominaLinea.objects.create(periodo=periodo, empleado=empleado, salario_base=Decimal(str(monto)))

        agregar_linea("NOM-CERRADA", date(2026, 3, 15), NominaPeriodo.ESTATUS_CERRADA, empleado_valido, 100)
        agregar_linea("NOM-PAGADA", date(2026, 3, 31), NominaPeriodo.ESTATUS_PAGADA, empleado_valido, 150)
        agregar_linea("NOM-BORRADOR", date(2026, 3, 20), NominaPeriodo.ESTATUS_BORRADOR, empleado_valido, 900)
        agregar_linea("NOM-OTRO-MES", date(2026, 2, 28), NominaPeriodo.ESTATUS_PAGADA, empleado_valido, 800)
        agregar_linea("NOM-OTRO-DEP", date(2026, 3, 10), NominaPeriodo.ESTATUS_PAGADA, empleado_otro_depto, 700)
        agregar_linea("NOM-OTRA-SUC", date(2026, 3, 11), NominaPeriodo.ESTATUS_PAGADA, empleado_otra_sucursal, 600)

        self.consolidar()

        linea.refresh_from_db()
        self.assertEqual(linea.monto_real, Decimal("250.00"))

    def test_venta_pos_normaliza_categoria_producto_y_filtra_sucursal(self):
        """VENTA_POS compara categoría/producto sin distinguir caso ni acentos."""
        rubro, linea = self.crear_linea(concepto="Bollo chocolate", sucursal=self.sucursal)
        ReglaFuenteRubro.objects.create(
            rubro=rubro,
            tipo_fuente=ReglaFuenteRubro.FUENTE_VENTA_POS,
            filtros={"categoria_pos": "BÓLLO", "producto_pos": "CHOCOLÁTE"},
        )
        branch = PointBranch.objects.create(external_id="POINT-GVE01", name="Centro", erp_branch=self.sucursal)
        otro_branch = PointBranch.objects.create(external_id="POINT-GVE02", name="Norte", erp_branch=self.otra_sucursal)

        def venta(branch_obj, fecha, categoria, producto, monto):
            PointSalesDailyProductFact.objects.create(
                branch=branch_obj,
                sale_date=fecha,
                sucursal_nombre=branch_obj.name,
                categoria=categoria,
                producto_nombre_historico=producto,
                total_venta=Decimal(str(monto)),
                total_venta_neta=Decimal(str(monto)) - Decimal("1"),
            )

        venta(branch, date(2026, 3, 2), "bollo", "chocolate", 100)
        venta(branch, date(2026, 3, 9), "BÓLLO", "CHOCOLÁTE", 50)
        venta(branch, date(2026, 2, 9), "BOLLO", "CHOCOLATE", 800)
        venta(branch, date(2026, 3, 10), "BOLLO", "VAINILLA", 700)
        venta(otro_branch, date(2026, 3, 11), "BOLLO", "CHOCOLATE", 600)

        self.consolidar()

        linea.refresh_from_db()
        self.assertEqual(linea.monto_real, Decimal("150.00"))

    def test_signo_resta_y_reglas_distintas_forman_fuente_ordenada(self):
        """Dos tipos de fuente se suman con signo y generan fuente alfabética."""
        rubro, linea = self.crear_linea(concepto="Combinado", sucursal=self.sucursal)
        self.crear_regla_gasto(rubro)
        ReglaFuenteRubro.objects.create(
            rubro=rubro,
            tipo_fuente=ReglaFuenteRubro.FUENTE_NOMINA,
            signo=-1,
            filtros={"campo_monto": "salario_base", "departamento": "VENTAS"},
        )
        self.crear_gasto("500")
        empleado = Empleado.objects.create(
            codigo="EMP-SIGNO", nombre="Empleado signo", departamento=Empleado.DEP_VENTAS, sucursal_ref=self.sucursal
        )
        nomina = NominaPeriodo.objects.create(
            folio="NOM-SIGNO", fecha_inicio=date(2026, 3, 1), fecha_fin=date(2026, 3, 31), estatus=NominaPeriodo.ESTATUS_CERRADA
        )
        NominaLinea.objects.create(periodo=nomina, empleado=empleado, salario_base=Decimal("120"))

        self.consolidar()

        linea.refresh_from_db()
        self.assertEqual(linea.monto_real, Decimal("380.00"))
        self.assertEqual(linea.fuente_real, "AUTO:GASTO_OPERATIVO+NOMINA")

    def test_linea_manual_nunca_cambia_y_se_cuenta_protegida(self):
        """Una captura MANUAL conserva monto, fuente y metadata."""
        rubro, linea = self.crear_linea(
            concepto="Manual", sucursal=self.sucursal, monto_real=Decimal("777"), fuente_real="MANUAL:johana"
        )
        linea.metadata = {"captura": "humana"}
        linea.save(update_fields=["metadata"])
        self.crear_regla_gasto(rubro)
        self.crear_gasto("100")

        summary = self.consolidar()

        linea.refresh_from_db()
        self.assertEqual(linea.monto_real, Decimal("777.00"))
        self.assertEqual(linea.fuente_real, "MANUAL:johana")
        self.assertEqual(linea.metadata, {"captura": "humana"})
        self.assertEqual(summary.protegidas_manual, 1)

    def test_segunda_consolidacion_es_idempotente(self):
        """La segunda corrida reconoce la línea AUTO sin cambios."""
        rubro, linea = self.crear_linea(concepto="Idempotente", sucursal=self.sucursal)
        self.crear_regla_gasto(rubro)
        self.crear_gasto("42")
        primera = self.consolidar()

        segunda = self.consolidar()

        linea.refresh_from_db()
        self.assertEqual(primera.actualizadas, 1)
        self.assertEqual(segunda.sin_cambio, 1)
        self.assertEqual(segunda.actualizadas, 0)
        self.assertEqual(linea.monto_real, Decimal("42.00"))

    def test_dry_run_no_persiste_y_detalla_el_cambio(self):
        """El modo dry-run calcula el detalle sin modificar la línea."""
        rubro, linea = self.crear_linea(concepto="Simulación", sucursal=self.sucursal)
        self.crear_regla_gasto(rubro)
        self.crear_gasto("63")

        summary = self.consolidar(dry_run=True)

        linea.refresh_from_db()
        self.assertIsNone(linea.monto_real)
        self.assertEqual(linea.fuente_real, "")
        self.assertEqual(summary.actualizadas, 1)
        self.assertEqual(len(summary.detalle), 1)
        self.assertEqual(summary.detalle[0]["nuevo"], "63.00")

    def test_regla_sin_datos_no_modifica_la_linea(self):
        """Sin filas fuente en el mes, la línea NO se toca (un retraso de
        Point/nómina no debe borrar el último real consolidado)."""
        rubro, linea = self.crear_linea(
            concepto="Sin datos",
            sucursal=self.sucursal,
            monto_real=Decimal("500.00"),
            fuente_real="AUTO:GASTO_OPERATIVO",
        )
        self.crear_regla_gasto(rubro)

        summary = self.consolidar()

        linea.refresh_from_db()
        self.assertEqual(linea.monto_real, Decimal("500.00"))
        self.assertEqual(linea.fuente_real, "AUTO:GASTO_OPERATIVO")
        self.assertEqual(summary.sin_datos_fuente, 1)
        self.assertEqual(summary.actualizadas, 0)
        # El valor retenido queda marcado visiblemente como fuente sin datos.
        self.assertTrue(linea.metadata["sin_datos_fuente"])
        self.assertIn("fuente_sin_datos_en", linea.metadata)
        # Y al volver datos de la fuente CON EL MISMO importe retenido, la
        # marca también se limpia (no queda un badge obsoleto por la ruta
        # "sin cambio").
        self.crear_gasto("500")
        self.consolidar()
        linea.refresh_from_db()
        self.assertEqual(linea.monto_real, Decimal("500.00"))
        self.assertFalse(linea.metadata.get("sin_datos_fuente"))
        self.assertNotIn("fuente_sin_datos_en", linea.metadata)

    def test_captura_manual_concurrente_no_se_pisa(self):
        """Si una usuaria captura entre la lectura y la escritura, el UPDATE
        condicional no coincide y la captura se conserva."""
        rubro, linea = self.crear_linea(concepto="Concurrente", sucursal=self.sucursal)
        self.crear_regla_gasto(rubro)
        self.crear_gasto("700")

        # Simula la carrera: la instancia en memoria tiene fuente_real="",
        # pero la base ya recibió una captura manual.
        LineaPresupuestoMensual.objects.filter(pk=linea.pk).update(
            monto_real=Decimal("123.45"), fuente_real="MANUAL:johana"
        )

        service = PresupuestoRealConsolidacionService()
        escrita = service._escribir_linea(linea, Decimal("700.00"), "AUTO:GASTO_OPERATIVO", {})

        self.assertFalse(escrita)
        linea.refresh_from_db()
        self.assertEqual(linea.monto_real, Decimal("123.45"))
        self.assertEqual(linea.fuente_real, "MANUAL:johana")

    def test_consolidacion_usa_consultas_acotadas(self):
        """Las fuentes se precargan agrupadas: el número de consultas no crece
        con el número de rubros."""
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        for i in range(12):
            rubro, _ = self.crear_linea(concepto=f"Rubro escala {i}", sucursal=self.sucursal)
            self.crear_regla_gasto(rubro)
        self.crear_gasto("100")

        with CaptureQueriesContext(connection) as ctx:
            self.consolidar(dry_run=True)
        # 12 rubros con regla: sin índices serían 12+ consultas de agregación.
        self.assertLess(len(ctx.captured_queries), 8)

    def test_rubro_sin_reglas_deja_linea_intacta(self):
        """Un rubro sin reglas se reporta y conserva todos sus valores."""
        _, linea = self.crear_linea(concepto="Sin regla", monto_real=Decimal("91"), fuente_real="fuente-anterior")
        linea.metadata = {"intacto": True}
        linea.save(update_fields=["metadata"])

        summary = self.consolidar()

        linea.refresh_from_db()
        self.assertEqual(summary.sin_regla, 1)
        self.assertEqual(linea.monto_real, Decimal("91.00"))
        self.assertEqual(linea.fuente_real, "fuente-anterior")
        self.assertEqual(linea.metadata, {"intacto": True})

    def test_migrar_fuentes_legadas_clasifica_auto_y_manual(self):
        """La migración convierte venta legada en AUTO y CAPEX en MANUAL protegido."""
        rubro_auto, linea_auto = self.crear_linea(
            concepto="Venta legada", monto_real=Decimal("10"), fuente_real="PROYECCIO_N_VENTAS_2026_AUTORIZADA"
        )
        rubro_manual, linea_manual = self.crear_linea(
            concepto="Capex legado", monto_real=Decimal("20"), fuente_real="CAPEX_GUAMUCHIL_CONFIRMADO"
        )
        self.crear_regla_gasto(rubro_auto)
        self.crear_regla_gasto(rubro_manual)

        resultado = migrar_fuentes_legadas()

        linea_auto.refresh_from_db()
        linea_manual.refresh_from_db()
        self.assertEqual(resultado["PROYECCIO_N_VENTAS_2026_AUTORIZADA"], 1)
        self.assertEqual(resultado["CAPEX_GUAMUCHIL_CONFIRMADO"], 1)
        self.assertEqual(linea_auto.fuente_real, "AUTO:LEGADO")
        self.assertEqual(linea_manual.fuente_real, "MANUAL:legado")
        # Con datos de fuente, el legado AUTO se re-escribe y el MANUAL queda protegido.
        self.crear_gasto("55")
        summary = self.consolidar()
        self.assertEqual(summary.actualizadas, 1)
        self.assertEqual(summary.protegidas_manual, 1)
        linea_auto.refresh_from_db()
        linea_manual.refresh_from_db()
        self.assertEqual(linea_auto.monto_real, Decimal("55.00"))
        self.assertEqual(linea_manual.monto_real, Decimal("20"))

    def test_seed_real_es_idempotente_respeta_admin_y_dry_run(self):
        """El seed real crea nómina/ventas, no duplica y respeta reglas ADMIN."""
        # El CSV real referencia RENTA_SUC; debe existir o el comando aborta.
        CategoriaGasto.objects.create(
            codigo="RENTA", nombre="Renta sucursal", capa_objetivo=CategoriaGasto.CAPA_EMPRESA
        )
        nomina = AreaPresupuesto.objects.create(nombre="Nómina seed", codigo="nomina")
        ventas = AreaPresupuesto.objects.create(nombre="Ventas seed", codigo="ventas")
        sueldo, _ = self.crear_linea(concepto="SUELDO", area=nomina)
        venta, _ = self.crear_linea(
            concepto="BOLLO · CHOCOLATE", area=ventas, sucursal=self.sucursal, tipo=RubroPresupuesto.TIPO_INGRESO
        )
        # Nombres POS reales para el matching difuso del seed.
        branch_seed = PointBranch.objects.create(external_id="SEED-BR", name="Centro", erp_branch=self.sucursal)
        PointSalesDailyProductFact.objects.create(
            branch=branch_seed,
            sale_date=self.periodo,
            sucursal_nombre="Centro",
            categoria="Bollo",
            producto_nombre_historico="Bollo Chocolate",
            total_venta=Decimal("10"),
            total_venta_neta=Decimal("9"),
        )
        sueldo_admin = RubroPresupuesto.objects.create(
            area=nomina, concepto="Sueldo", codigo_cuenta="ADMIN", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        ReglaFuenteRubro.objects.create(
            rubro=sueldo_admin,
            tipo_fuente=ReglaFuenteRubro.FUENTE_MANUAL,
            origen=ReglaFuenteRubro.ORIGEN_ADMIN,
        )

        call_command("seed_reglas_fuente_rubro", dry_run=True, stdout=StringIO())
        self.assertFalse(ReglaFuenteRubro.objects.filter(origen=ReglaFuenteRubro.ORIGEN_SEED).exists())

        call_command("seed_reglas_fuente_rubro", stdout=StringIO())
        total_seed = ReglaFuenteRubro.objects.filter(origen=ReglaFuenteRubro.ORIGEN_SEED).count()
        self.assertTrue(
            ReglaFuenteRubro.objects.filter(
                rubro=sueldo, origen=ReglaFuenteRubro.ORIGEN_SEED, tipo_fuente=ReglaFuenteRubro.FUENTE_NOMINA
            ).exists()
        )
        regla_venta = ReglaFuenteRubro.objects.get(rubro=venta, origen=ReglaFuenteRubro.ORIGEN_SEED)
        self.assertEqual(regla_venta.tipo_fuente, ReglaFuenteRubro.FUENTE_VENTA_POS)
        # El matching difuso asigna el nombre POS REAL, no el texto del rubro.
        self.assertEqual(regla_venta.filtros["productos_pos"], ["Bollo Chocolate"])
        self.assertEqual(venta.sucursal, self.sucursal)
        self.assertFalse(ReglaFuenteRubro.objects.filter(rubro=sueldo_admin, origen=ReglaFuenteRubro.ORIGEN_SEED).exists())

        call_command("seed_reglas_fuente_rubro", stdout=StringIO())
        self.assertEqual(ReglaFuenteRubro.objects.filter(origen=ReglaFuenteRubro.ORIGEN_SEED).count(), total_seed)
        self.assertEqual(ReglaFuenteRubro.objects.filter(rubro=sueldo_admin).count(), 1)

    def test_linea_auto_tiene_precedencia_sobre_resultado_empresa(self):
        """_line_actual usa el monto AUTO aunque exista un resultado mensual aplicable."""
        rubro, linea = self.crear_linea(
            concepto="Ventas", monto_real=Decimal("321.45"), fuente_real="AUTO:NOMINA", tipo=RubroPresupuesto.TIPO_INGRESO
        )
        rubro.metadata = {"actual_key": "ventas"}
        rubro.save(update_fields=["metadata"])
        EmpresaResultadoMensual.objects.create(periodo=self.periodo, venta_total=Decimal("9999"))
        linea.refresh_from_db()
        linea.rubro = rubro

        monto, fuente = PresupuestoMaestroService()._line_actual(
            linea, {"ventas": Decimal("9999")}, set()
        )

        self.assertEqual(monto, Decimal("321.45"))
        self.assertEqual(fuente, "AUTO:NOMINA")


class PresupuestoVsRealViewTests(TestCase):
    """Valida el tablero comparativo: RBAC, render y export."""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model

        User = get_user_model()
        cls.superuser = User.objects.create_superuser("dg_test", "dg@test.mx", "clave-test")
        cls.sin_permiso = User.objects.create_user("sin_permiso", "np@test.mx", "clave-test")

        cls.periodo = date(2026, 3, 1)
        cls.area = AreaPresupuesto.objects.create(nombre="Área tablero", codigo="tablero")
        cls.rubro = RubroPresupuesto.objects.create(
            area=cls.area, concepto="Concepto tablero", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        cls.linea = LineaPresupuestoMensual.objects.create(
            rubro=cls.rubro,
            periodo=cls.periodo,
            monto_presupuesto=Decimal("100.00"),
            monto_real=Decimal("80.00"),
            fuente_real="AUTO:NOMINA",
            metadata={"real_breakdown": [{"tipo_fuente": "NOMINA", "monto": "80.00"}]},
        )

    def test_requiere_permiso_de_reportes(self):
        """Un usuario sin acceso al módulo recibe 403."""
        self.client.force_login(self.sin_permiso)
        response = self.client.get("/reportes/presupuesto-vs-real/")
        self.assertEqual(response.status_code, 403)

    def test_render_muestra_detalle_y_fuente(self):
        """El tablero muestra el concepto, su varianza y el badge de fuente."""
        self.client.force_login(self.superuser)
        response = self.client.get("/reportes/presupuesto-vs-real/?year=2026&month=3")
        self.assertEqual(response.status_code, 200)
        contenido = response.content.decode()
        self.assertIn("Concepto tablero", contenido)
        self.assertIn("Automático · Nómina", contenido)
        detalle = response.context["detalle"]
        self.assertEqual(len(detalle), 1)
        self.assertEqual(detalle[0]["varianza"], Decimal("-20.00"))
        # Egreso gastando menos que presupuesto = verde
        self.assertEqual(detalle[0]["tone"], "success")

    def test_export_csv_incluye_encabezados_y_fila(self):
        """El export CSV trae encabezados y la línea del mes."""
        self.client.force_login(self.superuser)
        response = self.client.get("/reportes/presupuesto-vs-real/?year=2026&month=3&export=csv")
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response["Content-Type"])
        cuerpo = response.content.decode()
        self.assertIn("Concepto tablero", cuerpo)
        self.assertIn("Varianza %", cuerpo)

    def test_export_xlsx_responde_archivo(self):
        """El export XLSX responde un adjunto de Excel."""
        self.client.force_login(self.superuser)
        response = self.client.get("/reportes/presupuesto-vs-real/?year=2026&month=3&export=xlsx")
        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])
        self.assertIn("attachment", response["Content-Disposition"])


class PresupuestoRealFixesReviewTests(TestCase):
    """Cobertura de los hallazgos de la revisión adversarial."""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model

        cls.superuser = get_user_model().objects.create_superuser(
            "dg_fixes", "dgf@test.mx", "clave-test"
        )
        cls.periodo = date(2026, 3, 1)
        cls.area_nomina = AreaPresupuesto.objects.create(nombre="Nómina", codigo="nomina")
        cls.area_ventas = AreaPresupuesto.objects.create(nombre="Gastos", codigo="gastos-venta")

        rubro_nomina = RubroPresupuesto.objects.create(
            area=cls.area_nomina, concepto="SUELDO", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        LineaPresupuestoMensual.objects.create(
            rubro=rubro_nomina, periodo=cls.periodo,
            monto_presupuesto=Decimal("100.00"), monto_real=Decimal("90.00"),
            fuente_real="AUTO:NOMINA",
        )
        rubro_gasto = RubroPresupuesto.objects.create(
            area=cls.area_ventas, concepto="=SUMA(A1:A9)", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        LineaPresupuestoMensual.objects.create(
            rubro=rubro_gasto, periodo=cls.periodo,
            monto_presupuesto=Decimal("50.00"), monto_real=Decimal("40.00"),
            fuente_real="AUTO:GASTO_OPERATIVO",
        )

    def test_kpi_global_excluye_area_nomina(self):
        """El área Nómina no se suma a los KPI globales (doble conteo)."""
        self.client.force_login(self.superuser)
        response = self.client.get("/reportes/presupuesto-vs-real/?year=2026&month=3")
        kpis = response.context["kpis"]
        self.assertEqual(kpis["presupuesto"], Decimal("50.00"))
        self.assertEqual(kpis["real"], Decimal("40.00"))
        # Con el área nómina seleccionada sí se muestra su propio total.
        response = self.client.get("/reportes/presupuesto-vs-real/?year=2026&month=3&area=nomina")
        self.assertEqual(response.context["kpis"]["presupuesto"], Decimal("100.00"))

    def test_export_neutraliza_formulas(self):
        """Un concepto que empieza con '=' se exporta neutralizado."""
        self.client.force_login(self.superuser)
        response = self.client.get("/reportes/presupuesto-vs-real/?year=2026&month=3&export=csv")
        cuerpo = response.content.decode()
        self.assertIn("'=SUMA(A1:A9)", cuerpo)
        self.assertNotIn("\n=SUMA", cuerpo.replace("\r", ""))

    def test_seed_elimina_reglas_obsoletas(self):
        """Una regla SEED cuyo rubro salió del mapeo se elimina al re-correr."""
        CategoriaGasto.objects.create(
            codigo="RENTA", nombre="Renta sucursal", capa_objetivo=CategoriaGasto.CAPA_EMPRESA
        )
        rubro_viejo = RubroPresupuesto.objects.create(
            area=self.area_ventas, concepto="Concepto retirado", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        ReglaFuenteRubro.objects.create(
            rubro=rubro_viejo,
            tipo_fuente=ReglaFuenteRubro.FUENTE_NOMINA,
            origen=ReglaFuenteRubro.ORIGEN_SEED,
            filtros={"campo_monto": "salario_base"},
        )
        salida = StringIO()
        call_command("seed_reglas_fuente_rubro", stdout=salida)
        self.assertFalse(ReglaFuenteRubro.objects.filter(rubro=rubro_viejo).exists())
        self.assertIn("seed obsoletas eliminadas: 1", salida.getvalue())

    def test_categoria_inexistente_aborta_sin_escribir(self):
        """Una categoria_gasto inválida en el CSV aborta el comando completo;
        las reglas SEED previas se conservan (no hay borrado degradado)."""
        import csv as csv_mod
        import tempfile

        from django.core.management.base import CommandError

        rubro = RubroPresupuesto.objects.create(
            area=self.area_ventas, concepto="Con regla previa", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        regla_previa = ReglaFuenteRubro.objects.create(
            rubro=rubro,
            tipo_fuente=ReglaFuenteRubro.FUENTE_NOMINA,
            origen=ReglaFuenteRubro.ORIGEN_SEED,
            filtros={"campo_monto": "salario_base"},
        )
        with tempfile.NamedTemporaryFile("w", suffix=".csv", delete=False, encoding="utf-8") as tmp:
            writer = csv_mod.DictWriter(
                tmp, fieldnames=["area", "concepto", "tipo_fuente", "categoria_gasto", "filtros", "notas"]
            )
            writer.writeheader()
            writer.writerow(
                {
                    "area": "gastos-venta",
                    "concepto": "Con regla previa",
                    "tipo_fuente": "GASTO_OPERATIVO",
                    "categoria_gasto": "NO_EXISTE_XYZ",
                    "filtros": "",
                    "notas": "",
                }
            )
            ruta = tmp.name

        with self.assertRaises(CommandError):
            call_command("seed_reglas_fuente_rubro", csv=ruta, stdout=StringIO())
        self.assertTrue(ReglaFuenteRubro.objects.filter(pk=regla_previa.pk).exists())


class VentasPosMatchingTests(TestCase):
    """Matching difuso rubro de Ventas → nombres POS reales (casos de producción)."""

    @classmethod
    def setUpTestData(cls):
        # El CSV real referencia RENTA_SUC; debe existir o el comando aborta.
        CategoriaGasto.objects.create(
            codigo="RENTA", nombre="Renta sucursal", capa_objetivo=CategoriaGasto.CAPA_EMPRESA
        )
        cls.ventas = AreaPresupuesto.objects.create(nombre="Ventas", codigo="ventas")
        cls.branch = PointBranch.objects.create(external_id="MATCH-BR", name="Centro")

        def fact(categoria, producto):
            PointSalesDailyProductFact.objects.create(
                branch=cls.branch,
                sale_date=date(2026, 5, 3),
                sucursal_nombre="Centro",
                categoria=categoria,
                producto_nombre_historico=producto,
                total_venta=Decimal("10"),
                total_venta_neta=Decimal("9"),
            )

        fact("Pastel Mediano", "Pastel de 3 Pecados Mediano")
        fact("Pastel Mediano", "Pastel de Snickers Mediano")
        fact("Rebanada", "Pastel de 3 Pecados R")
        fact("TE", "Té helado 500ml")
        fact("Galletas", "Bolitas de Nuez 10 PZ")

    def _rubro(self, concepto):
        rubro = RubroPresupuesto.objects.create(
            area=self.ventas, concepto=concepto, tipo=RubroPresupuesto.TIPO_INGRESO
        )
        LineaPresupuestoMensual.objects.create(
            rubro=rubro, periodo=date(2026, 5, 1), monto_presupuesto=Decimal("1")
        )
        return rubro

    def _filtros_de(self, rubro):
        call_command("seed_reglas_fuente_rubro", stdout=StringIO())
        return ReglaFuenteRubro.objects.get(rubro=rubro, origen=ReglaFuenteRubro.ORIGEN_SEED).filtros

    def test_producto_con_orden_y_preposiciones_distintas(self):
        """'PASTEL MEDIANO · 3 PECADOS' cruza con 'Pastel de 3 Pecados Mediano'."""
        rubro = self._rubro("PASTEL MEDIANO · 3 PECADOS")
        filtros = self._filtros_de(rubro)
        self.assertEqual(filtros["productos_pos"], ["Pastel de 3 Pecados Mediano"])

    def test_apostrofe_y_abreviatura(self):
        """SNICKER'S cruza con Snickers; el sufijo R se expande a rebanada."""
        rubro_snickers = self._rubro("PASTEL MEDIANO · SNICKER'S")
        rubro_rebanada = self._rubro("PASTEL REBANADAS · 3 PECADOS")
        call_command("seed_reglas_fuente_rubro", stdout=StringIO())
        filtros_s = ReglaFuenteRubro.objects.get(rubro=rubro_snickers, origen="SEED").filtros
        filtros_r = ReglaFuenteRubro.objects.get(rubro=rubro_rebanada, origen="SEED").filtros
        self.assertEqual(filtros_s["productos_pos"], ["Pastel de Snickers Mediano"])
        self.assertEqual(filtros_r["productos_pos"], ["Pastel de 3 Pecados R"])

    def test_categoria_pos_completa(self):
        """'BEBIDAS/OTROS · TE' cruza con la categoría POS 'TE' completa."""
        rubro = self._rubro("BEBIDAS/OTROS · TE")
        filtros = self._filtros_de(rubro)
        self.assertEqual(filtros.get("categoria_pos"), "TE")
        self.assertNotIn("productos_pos", filtros)

    def test_sin_match_queda_reportado_y_sin_asignacion(self):
        """Un concepto sin equivalente POS ni override queda sin asignación."""
        rubro = self._rubro("BEBIDAS/OTROS · KOMBUCHA")
        filtros = self._filtros_de(rubro)
        self.assertNotIn("productos_pos", filtros)
        self.assertNotIn("categoria_pos", filtros)

    def test_conflicto_de_producto_gana_el_mejor_score(self):
        """Dos rubros no pueden reclamar el mismo producto POS."""
        exacto = self._rubro("GALLETA · BOLITA DE NUEZ (10PZ)")
        parecido = self._rubro("GALLETA · BOLITAS DE NUEZ")
        call_command("seed_reglas_fuente_rubro", stdout=StringIO())
        con_producto = [
            r for r in ReglaFuenteRubro.objects.filter(rubro__in=[exacto, parecido])
            if r.filtros.get("productos_pos")
        ]
        self.assertEqual(len(con_producto), 1)


class VentasUnidadesTests(TestCase):
    """Comparativo de ventas por unidades × precio actual (regla de dirección)."""

    @classmethod
    def setUpTestData(cls):
        from recetas.models import PronosticoVenta, Receta

        cls.receta = Receta.objects.create(nombre="Bollo Chocolate", hash_contenido="test-vu-bollo")
        cls.sin_venta = Receta.objects.create(nombre="Rosca Dulce de Leche", hash_contenido="test-vu-rosca")
        PronosticoVenta.objects.create(receta=cls.receta, periodo="2026-05", cantidad=Decimal("100"), fuente="PRESUPUESTO_2026")
        PronosticoVenta.objects.create(receta=cls.sin_venta, periodo="2026-05", cantidad=Decimal("40"), fuente="PRESUPUESTO_2026")

        branch = PointBranch.objects.create(external_id="VU-BR", name="Centro")
        for day, qty, venta in [(3, 30, 300), (10, 50, 500)]:
            PointSalesDailyProductFact.objects.create(
                branch=branch,
                sale_date=date(2026, 5, day),
                sucursal_nombre="Centro",
                categoria="Bollo",
                producto_nombre_historico="Bollo Chocolate",
                receta=cls.receta,
                total_cantidad=Decimal(qty),
                total_venta=Decimal(venta),
                total_venta_neta=Decimal(venta),
            )

    def test_unidades_cumplimiento_e_importe_a_precio_actual(self):
        """80 de 100 unidades = 80% y el $ proyectado usa el ASP reciente ($10)."""
        from reportes.services_ventas_unidades import comparativo_ventas_unidades

        resultado = comparativo_ventas_unidades(date(2026, 5, 1), hoy=date(2026, 5, 31))
        fila = next(f for f in resultado["filas"] if f["receta_id"] == self.receta.id)
        self.assertEqual(fila["unidades_proyectadas"], Decimal("100"))
        self.assertEqual(fila["unidades_reales"], Decimal("80"))
        self.assertEqual(fila["cumplimiento_pct"], Decimal("80.0"))
        self.assertEqual(fila["precio_actual"], Decimal("10.00"))
        self.assertEqual(fila["importe_proyectado"], Decimal("1000.00"))
        self.assertEqual(fila["importe_real"], Decimal("800"))
        self.assertEqual(fila["varianza"], Decimal("-200.00"))

    def test_producto_sin_venta_ni_precio_queda_marcado(self):
        """Sin ventas recientes ni precio de lista: fila sin precio, contada."""
        from reportes.services_ventas_unidades import comparativo_ventas_unidades

        resultado = comparativo_ventas_unidades(date(2026, 5, 1), hoy=date(2026, 5, 31))
        fila = next(f for f in resultado["filas"] if f["receta_id"] == self.sin_venta.id)
        self.assertIsNone(fila["precio_actual"])
        self.assertIsNone(fila["importe_proyectado"])
        self.assertEqual(resultado["totales"]["sin_precio"], 1)


class ImportUnidadesProyeccionTests(TestCase):
    """El reimport de la proyección guarda las CANTIDADES en PronosticoVenta."""

    def _xlsx_proyeccion(self):
        import tempfile

        from openpyxl import Workbook

        wb = Workbook()
        hoja = wb.active
        hoja.title = "GENERAL"
        # Fila de meses + encabezados CANT/VENTA por bloque (proyección 2026).
        hoja.append(["", "ENERO", "", "", "FEBRERO", "", "", "MARZO", "", ""])
        hoja.append(["", "PROYECCIÓN 2026", "", "", "PROYECCIÓN 2026", "", "", "PROYECCIÓN 2026", "", ""])
        hoja.append(["CONCEPTO", "CANT", "VENTA", "", "CANT", "VENTA", "", "CANT", "VENTA", ""])
        hoja.append(["BOLLO", None, None, None, None, None, None, None, None, None])
        hoja.append(["CHOCOLATE", 120, 3600, None, 150, 4500, None, 130, 3900, None])
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        return tmp.name

    def test_reimport_guarda_unidades_y_respeta_manual(self):
        from recetas.models import PronosticoVenta, Receta

        from reportes.services_presupuesto_maestro import PresupuestoMaestroImportService

        receta = Receta.objects.create(nombre="Bollo Chocolate", hash_contenido="test-imp-bollo")
        # Pronóstico manual previo de enero: NO debe pisarse.
        PronosticoVenta.objects.create(receta=receta, periodo="2026-01", cantidad=Decimal("999"), fuente="MANUAL")

        resumen = PresupuestoMaestroImportService().reimport_sales_projection(
            archivo=self._xlsx_proyeccion(), year=2026
        )

        enero = PronosticoVenta.objects.get(receta=receta, periodo="2026-01")
        febrero = PronosticoVenta.objects.get(receta=receta, periodo="2026-02")
        marzo = PronosticoVenta.objects.get(receta=receta, periodo="2026-03")
        self.assertEqual(enero.cantidad, Decimal("999"))
        self.assertEqual(enero.fuente, "MANUAL")
        self.assertEqual(febrero.cantidad, Decimal("150"))
        self.assertEqual(febrero.fuente, "PRESUPUESTO_2026")
        self.assertEqual(marzo.cantidad, Decimal("130"))
        self.assertEqual(resumen.unidades_upsertadas, 2)


class CapturaPorAreaTests(TestCase):
    """Pantalla de captura distribuida: RBAC por área e invariantes de escritura."""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model

        User = get_user_model()
        cls.dg = User.objects.create_superuser("dg_captura", "dgc@test.mx", "clave-test")
        cls.jefa_logistica = User.objects.create_user("jefa_logistica", "jl@test.mx", "clave-test")
        cls.sin_area = User.objects.create_user("sin_area_cap", "sa@test.mx", "clave-test")

        cls.periodo = date(2026, 6, 1)
        cls.area_logistica = AreaPresupuesto.objects.create(nombre="Logística", codigo="logistica")
        cls.area_admin = AreaPresupuesto.objects.create(nombre="Administración", codigo="administracion")
        AreaPresupuestoResponsable = __import__(
            "reportes.models", fromlist=["AreaPresupuestoResponsable"]
        ).AreaPresupuestoResponsable
        AreaPresupuestoResponsable.objects.create(area=cls.area_logistica, usuario=cls.jefa_logistica)

        def linea(area, concepto, **kwargs):
            rubro = RubroPresupuesto.objects.create(
                area=area, concepto=concepto, tipo=RubroPresupuesto.TIPO_EGRESO
            )
            return LineaPresupuestoMensual.objects.create(
                rubro=rubro, periodo=cls.periodo, monto_presupuesto=Decimal("100"), **kwargs
            )

        cls.linea_manual = linea(cls.area_logistica, "Diesel")
        cls.linea_otra_area = linea(cls.area_admin, "ISR")
        cls.linea_auto = linea(cls.area_logistica, "Sueldo logística")
        ReglaFuenteRubro.objects.create(
            rubro=cls.linea_auto.rubro,
            tipo_fuente=ReglaFuenteRubro.FUENTE_NOMINA,
            filtros={"campo_monto": "salario_base", "departamento": "LOGISTICA"},
        )

    URL = "/reportes/presupuesto-real/captura/"
    URL_GUARDAR = "/reportes/presupuesto-real/captura/guardar/"

    def test_usuario_sin_area_no_entra(self):
        self.client.force_login(self.sin_area)
        self.assertEqual(self.client.get(self.URL).status_code, 403)

    def test_jefa_solo_ve_su_area(self):
        self.client.force_login(self.jefa_logistica)
        response = self.client.get(f"{self.URL}?year=2026&month=6")
        self.assertEqual(response.status_code, 200)
        codigos = [a.codigo for a in response.context["areas"]]
        self.assertEqual(codigos, ["logistica"])
        conceptos = [f["linea"].rubro.concepto for f in response.context["filas"]]
        self.assertIn("Diesel", conceptos)
        self.assertNotIn("ISR", conceptos)

    def test_guardar_escribe_manual_con_historial(self):
        self.client.force_login(self.jefa_logistica)
        response = self.client.post(
            self.URL_GUARDAR,
            {"linea_id": self.linea_manual.id, "monto": "1,250.50"},
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["toast"]["type"], "success")
        self.linea_manual.refresh_from_db()
        self.assertEqual(self.linea_manual.monto_real, Decimal("1250.50"))
        self.assertEqual(self.linea_manual.fuente_real, "MANUAL:jefa_logistica")
        self.assertEqual(len(self.linea_manual.metadata["capturas"]), 1)
        # Y la consolidación automática NO la pisa después.
        service = PresupuestoRealConsolidacionService()
        summary = service.consolidar(periodo=self.periodo)
        self.linea_manual.refresh_from_db()
        self.assertEqual(self.linea_manual.monto_real, Decimal("1250.50"))

    def test_no_captura_linea_de_otra_area_ni_automatica(self):
        self.client.force_login(self.jefa_logistica)
        otra = self.client.post(
            self.URL_GUARDAR,
            {"linea_id": self.linea_otra_area.id, "monto": "10"},
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(otra.status_code, 403)
        auto = self.client.post(
            self.URL_GUARDAR,
            {"linea_id": self.linea_auto.id, "monto": "10"},
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(auto.status_code, 409)
        self.linea_auto.refresh_from_db()
        self.assertIsNone(self.linea_auto.monto_real)

    def test_monto_invalido_rechazado(self):
        self.client.force_login(self.jefa_logistica)
        response = self.client.post(
            self.URL_GUARDAR,
            {"linea_id": self.linea_manual.id, "monto": "abc"},
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(response.status_code, 400)

    def test_post_tradicional_redirige_con_fragmento(self):
        self.client.force_login(self.dg)
        response = self.client.post(
            self.URL_GUARDAR,
            {
                "linea_id": self.linea_manual.id,
                "monto": "500",
                "return_to": f"{self.URL}?year=2026&month=6&area=logistica",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn(f"#linea-{self.linea_manual.id}", response["Location"])


class EndurecimientoAuditoriaTests(TestCase):
    """Fixes de la auditoría de arquitectura: importador, matching y RBAC."""

    def _csv_ventas(self, concepto="Rubro import", real_enero=None):
        import csv as csv_mod
        import tempfile

        campos = ["concepto", "tipo", "sucursal", "enero", "febrero", "marzo", "abril", "mayo",
                  "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
        fila = {c: "0" for c in campos}
        fila.update({"concepto": concepto, "tipo": "EGRESO", "sucursal": "", "enero": "100.00"})
        if real_enero is not None:
            campos.append("enero_real")
            fila["enero_real"] = real_enero
        tmp = tempfile.NamedTemporaryFile("w", newline="", suffix=".csv", delete=False, encoding="utf-8")
        writer = csv_mod.DictWriter(tmp, fieldnames=campos)
        writer.writeheader()
        writer.writerow(fila)
        tmp.close()
        return tmp.name

    def test_import_no_pisa_captura_manual_ni_metadata(self):
        """Re-importar presupuesto conserva el real MANUAL y su historial."""
        from reportes.services_presupuesto_maestro import PresupuestoMaestroImportService

        service = PresupuestoMaestroImportService()
        service.import_file(archivo=self._csv_ventas(), area_code="administracion", version="ORIGINAL", year=2026)
        linea = LineaPresupuestoMensual.objects.get(periodo=date(2026, 1, 1), rubro__concepto="Rubro import")
        linea.monto_real = Decimal("777.77")
        linea.fuente_real = "MANUAL:paula.lugo"
        linea.metadata = {**linea.metadata, "capturas": [{"usuario": "paula.lugo", "monto": "777.77"}]}
        linea.save()

        # Re-import con columna _real que intenta traer otro valor.
        service.import_file(
            archivo=self._csv_ventas(real_enero="999.99"),
            area_code="administracion", version="ORIGINAL", year=2026,
        )
        linea.refresh_from_db()
        self.assertEqual(linea.monto_real, Decimal("777.77"))
        self.assertEqual(linea.fuente_real, "MANUAL:paula.lugo")
        self.assertEqual(linea.metadata["capturas"][0]["usuario"], "paula.lugo")
        # Y el presupuesto sí se actualizó (ese es el trabajo del import).
        self.assertEqual(linea.monto_presupuesto, Decimal("100.00"))

    def test_import_real_en_linea_libre_usa_namespace_legado(self):
        """Columna _real en línea sin captura entra como AUTO:LEGADO (re-escribible)."""
        from reportes.services_presupuesto_maestro import PresupuestoMaestroImportService

        PresupuestoMaestroImportService().import_file(
            archivo=self._csv_ventas(concepto="Rubro libre", real_enero="55.00"),
            area_code="administracion", version="ORIGINAL", year=2026,
        )
        linea = LineaPresupuestoMensual.objects.get(periodo=date(2026, 1, 1), rubro__concepto="Rubro libre")
        self.assertEqual(linea.monto_real, Decimal("55.00"))
        self.assertEqual(linea.fuente_real, "AUTO:LEGADO")

    def test_clear_first_bloqueado_con_capturas_manuales(self):
        """reimport_sales_projection(clear_first=True) no destruye capturas."""
        from reportes.services_presupuesto_maestro import (
            PresupuestoMaestroImportService,
            ensure_master_budget_areas,
        )

        areas = ensure_master_budget_areas()
        rubro = RubroPresupuesto.objects.create(
            area=areas["ventas"], concepto="Con captura", tipo=RubroPresupuesto.TIPO_INGRESO
        )
        LineaPresupuestoMensual.objects.create(
            rubro=rubro, periodo=date(2026, 1, 1), monto_presupuesto=Decimal("1"),
            monto_real=Decimal("10"), fuente_real="MANUAL:johana",
        )
        with self.assertRaises(ValueError):
            PresupuestoMaestroImportService().reimport_sales_projection(
                archivo=self._csv_ventas(), year=2026, clear_first=True
            )
        self.assertTrue(LineaPresupuestoMensual.objects.filter(fuente_real="MANUAL:johana").exists())

    def test_clear_first_bloqueado_con_auto_y_reglas_admin(self):
        """clear_first tampoco destruye consolidados AUTO ni reglas ADMIN."""
        from reportes.services_presupuesto_maestro import (
            PresupuestoMaestroImportService,
            ensure_master_budget_areas,
        )

        areas = ensure_master_budget_areas()
        rubro = RubroPresupuesto.objects.create(
            area=areas["ventas"], concepto="Consolidado auto", tipo=RubroPresupuesto.TIPO_INGRESO
        )
        LineaPresupuestoMensual.objects.create(
            rubro=rubro, periodo=date(2026, 1, 1), monto_presupuesto=Decimal("1"),
            monto_real=Decimal("10"), fuente_real="AUTO:VENTA_POS",
        )
        with self.assertRaises(ValueError):
            PresupuestoMaestroImportService().reimport_sales_projection(
                archivo=self._csv_ventas(), year=2026, clear_first=True
            )

        # Solo regla ADMIN, sin reales: también bloquea.
        LineaPresupuestoMensual.objects.filter(rubro=rubro).update(monto_real=None, fuente_real="")
        ReglaFuenteRubro.objects.create(
            rubro=rubro, tipo_fuente=ReglaFuenteRubro.FUENTE_MANUAL,
            origen=ReglaFuenteRubro.ORIGEN_ADMIN,
        )
        with self.assertRaises(ValueError):
            PresupuestoMaestroImportService().reimport_sales_projection(
                archivo=self._csv_ventas(), year=2026, clear_first=True
            )
        self.assertTrue(RubroPresupuesto.objects.filter(pk=rubro.pk).exists())

    def test_matching_no_regala_toppings_ni_categorias_solapadas(self):
        """token_sort no asigna superconjuntos y la categoría solapada se anula."""
        CategoriaGasto.objects.create(
            codigo="RENTA", nombre="Renta sucursal", capa_objetivo=CategoriaGasto.CAPA_EMPRESA
        )
        ventas = AreaPresupuesto.objects.create(nombre="Ventas", codigo="ventas")
        branch = PointBranch.objects.create(external_id="HARD-BR", name="Centro")

        def fact(categoria, producto):
            PointSalesDailyProductFact.objects.create(
                branch=branch, sale_date=date(2026, 5, 3), sucursal_nombre="Centro",
                categoria=categoria, producto_nombre_historico=producto,
                total_venta=Decimal("10"), total_venta_neta=Decimal("9"),
            )

        fact("Pastel Chico", "Pastel de Crunch Chico")
        fact("Pastel Chico", "TOPPING CRUNCH C")
        fact("Rebanada", "Pastel de 3 Pecados R")

        def rubro(concepto):
            r = RubroPresupuesto.objects.create(area=ventas, concepto=concepto, tipo=RubroPresupuesto.TIPO_INGRESO)
            LineaPresupuestoMensual.objects.create(rubro=r, periodo=date(2026, 5, 1), monto_presupuesto=Decimal("1"))
            return r

        r_crunch = rubro("PASTEL CHICO · CRUNCH")
        r_rebanada = rubro("PASTEL REBANADAS · 3 PECADOS")
        r_fresa = rubro("CHEESECAKE · FRESA R")  # antes recibía "Rebanada" completa

        call_command("seed_reglas_fuente_rubro", stdout=StringIO())

        filtros_crunch = ReglaFuenteRubro.objects.get(rubro=r_crunch).filtros
        self.assertEqual(filtros_crunch["productos_pos"], ["Pastel de Crunch Chico"])  # sin topping
        filtros_rebanada = ReglaFuenteRubro.objects.get(rubro=r_rebanada).filtros
        self.assertEqual(filtros_rebanada["productos_pos"], ["Pastel de 3 Pecados R"])
        filtros_fresa = ReglaFuenteRubro.objects.get(rubro=r_fresa).filtros
        # La categoría "Rebanada" ya tiene productos asignados a otro rubro:
        # no puede quedar como categoría completa (doble conteo).
        self.assertNotIn("categoria_pos", filtros_fresa)
        self.assertNotIn("productos_pos", filtros_fresa)

    def test_post_presupuesto_maestro_requiere_manage(self):
        """Un usuario sin nivel de administración no puede escribir por POST."""
        from django.contrib.auth import get_user_model

        comun = get_user_model().objects.create_user("solo_lectura_x", "sl@test.mx", "clave-test")
        self.client.force_login(comun)
        r1 = self.client.post("/reportes/presupuesto-maestro/", {"action": "add_rubro"})
        self.assertEqual(r1.status_code, 403)
        r2 = self.client.post("/reportes/presupuestos/importar/", {})
        self.assertEqual(r2.status_code, 403)


class LimpiezaSinAsignacionTests(TestCase):
    """Reales AUTO huérfanos de una asignación anulada vuelven a pendiente."""

    def test_limpia_auto_de_regla_sin_asignacion_y_respeta_manual(self):
        from reportes.services_presupuesto_real import limpiar_reales_sin_asignacion

        area = AreaPresupuesto.objects.create(nombre="Ventas", codigo="ventas")

        def caso(concepto, fuente, filtros):
            rubro = RubroPresupuesto.objects.create(
                area=area, concepto=concepto, tipo=RubroPresupuesto.TIPO_INGRESO
            )
            ReglaFuenteRubro.objects.create(
                rubro=rubro, tipo_fuente=ReglaFuenteRubro.FUENTE_VENTA_POS, filtros=filtros
            )
            return LineaPresupuestoMensual.objects.create(
                rubro=rubro, periodo=date(2026, 5, 1), monto_presupuesto=Decimal("1"),
                monto_real=Decimal("999"), fuente_real=fuente,
            )

        huerfana = caso("Sin asignación", "AUTO:VENTA_POS", {"campo_monto": "total_venta"})
        asignada = caso("Con asignación", "AUTO:VENTA_POS", {"productos_pos": ["Bollo Lotus"]})
        manual = caso("Capturada", "MANUAL:johana", {"campo_monto": "total_venta"})

        limpiadas = limpiar_reales_sin_asignacion()

        huerfana.refresh_from_db(); asignada.refresh_from_db(); manual.refresh_from_db()
        self.assertEqual(limpiadas, 1)
        self.assertIsNone(huerfana.monto_real)
        self.assertEqual(huerfana.fuente_real, "")
        self.assertIn("limpiado_sin_asignacion_en", huerfana.metadata)
        self.assertEqual(asignada.monto_real, Decimal("999"))
        self.assertEqual(manual.monto_real, Decimal("999"))


class MatchingProductoSoloTests(TestCase):
    """El score toma el mejor entre producto-solo y categoría+producto."""

    def test_categoria_ruidosa_no_bloquea_match_exacto(self):
        """'GALLETA · LOTUS' cruza con 'Galleta Lotus' aunque la categoría POS
        sea 'Galletas' (plural); y el topping sigue sin regalarse."""
        CategoriaGasto.objects.create(
            codigo="RENTA", nombre="Renta sucursal", capa_objetivo=CategoriaGasto.CAPA_EMPRESA
        )
        area = AreaPresupuesto.objects.create(nombre="Ventas", codigo="ventas")
        branch = PointBranch.objects.create(external_id="MPS-BR", name="Centro")
        for cat, prod in [("Galletas", "Galleta Lotus"), ("Pastel Chico", "TOPPING CRUNCH C"),
                          ("Pastel Chico", "Pastel de Crunch Chico")]:
            PointSalesDailyProductFact.objects.create(
                branch=branch, sale_date=date(2026, 5, 3), sucursal_nombre="Centro",
                categoria=cat, producto_nombre_historico=prod,
                total_venta=Decimal("10"), total_venta_neta=Decimal("9"),
            )

        def rubro(concepto):
            r = RubroPresupuesto.objects.create(area=area, concepto=concepto, tipo=RubroPresupuesto.TIPO_INGRESO)
            LineaPresupuestoMensual.objects.create(rubro=r, periodo=date(2026, 5, 1), monto_presupuesto=Decimal("1"))
            return r

        r_lotus = rubro("GALLETA · LOTUS")
        r_crunch = rubro("PASTEL CHICO · CRUNCH")
        call_command("seed_reglas_fuente_rubro", stdout=StringIO())

        self.assertEqual(
            ReglaFuenteRubro.objects.get(rubro=r_lotus).filtros["productos_pos"], ["Galleta Lotus"]
        )
        self.assertEqual(
            ReglaFuenteRubro.objects.get(rubro=r_crunch).filtros["productos_pos"], ["Pastel de Crunch Chico"]
        )


class MatchingRefinadoTests(TestCase):
    """Apóstrofes, desempate por producto y overrides de CSV en ventas."""

    @classmethod
    def setUpTestData(cls):
        CategoriaGasto.objects.create(
            codigo="RENTA", nombre="Renta sucursal", capa_objetivo=CategoriaGasto.CAPA_EMPRESA
        )
        cls.area = AreaPresupuesto.objects.create(nombre="Ventas", codigo="ventas")
        cls.branch = PointBranch.objects.create(external_id="REF-BR", name="Centro")
        for cat, prod in [
            ("Pastel Chico", "Pastel de Snickers Chico"),
            ("Pastel Grande", "Pastel de Fresas Con Crema Grande"),
            ("Vasos Grande", "Vaso Fresas con Crema Grande"),
            ("Coca-cola", "COCA-COLA 450 ML"),
        ]:
            PointSalesDailyProductFact.objects.create(
                branch=cls.branch, sale_date=date(2026, 5, 3), sucursal_nombre="Centro",
                categoria=cat, producto_nombre_historico=prod,
                total_venta=Decimal("10"), total_venta_neta=Decimal("9"),
            )

    def _rubro(self, concepto):
        r = RubroPresupuesto.objects.create(
            area=self.area, concepto=concepto, tipo=RubroPresupuesto.TIPO_INGRESO
        )
        LineaPresupuestoMensual.objects.create(
            rubro=r, periodo=date(2026, 5, 1), monto_presupuesto=Decimal("1")
        )
        return r

    def test_apostrofe_no_rompe_el_match(self):
        """SNICKER'S cruza con Snickers (apóstrofe eliminado, no separado)."""
        rubro = self._rubro("PASTEL CHICO · SNICKER'S")
        call_command("seed_reglas_fuente_rubro", stdout=StringIO())
        filtros = ReglaFuenteRubro.objects.get(rubro=rubro).filtros
        self.assertEqual(filtros["productos_pos"], ["Pastel de Snickers Chico"])

    def test_desempate_por_producto_no_por_rubro(self):
        """El vaso preparado gana SU producto aunque el rubro de pastel tenga
        mejor score global en el suyo propio."""
        r_pastel = self._rubro("PASTEL GRANDE · FRESAS CON CREMA")
        r_vaso = self._rubro("VASO PREPARADO · FRESAS CON CREMA GDE")
        call_command("seed_reglas_fuente_rubro", stdout=StringIO())
        filtros_pastel = ReglaFuenteRubro.objects.get(rubro=r_pastel).filtros
        filtros_vaso = ReglaFuenteRubro.objects.get(rubro=r_vaso).filtros
        self.assertIn("Pastel de Fresas Con Crema Grande", filtros_pastel.get("productos_pos", []))
        self.assertNotIn("Vaso Fresas con Crema Grande", filtros_pastel.get("productos_pos", []))
        self.assertEqual(filtros_vaso.get("productos_pos"), ["Vaso Fresas con Crema Grande"])

    def test_override_de_csv_manda_sobre_la_autoasignacion(self):
        """El rubro con override en el CSV recibe SOLO las reglas del CSV."""
        rubro = self._rubro("Bebidas")
        call_command("seed_reglas_fuente_rubro", stdout=StringIO())
        reglas = list(ReglaFuenteRubro.objects.filter(rubro=rubro))
        self.assertEqual(len(reglas), 3)
        self.assertEqual(
            {r.filtros.get("categoria_pos") for r in reglas}, {"Coca-cola", "TE", "Café"}
        )


class FueraProyeccionTests(TestCase):
    """Productos con venta real sin proyección aparecen en su propia sección."""

    def test_producto_no_proyectado_se_reporta(self):
        from recetas.models import PronosticoVenta, Receta

        from reportes.services_ventas_unidades import comparativo_ventas_unidades

        proyectada = Receta.objects.create(nombre="Bollo Lotus", hash_contenido="fp-lotus")
        temporada = Receta.objects.create(nombre="Rosca de Reyes", hash_contenido="fp-rosca")
        PronosticoVenta.objects.create(
            receta=proyectada, periodo="2026-05", cantidad=Decimal("10"), fuente="PRESUPUESTO_2026"
        )
        branch = PointBranch.objects.create(external_id="FP-BR", name="Centro")

        def venta(receta, producto, monto, qty):
            PointSalesDailyProductFact.objects.create(
                branch=branch, sale_date=date(2026, 5, 9), sucursal_nombre="Centro",
                categoria="X", producto_nombre_historico=producto, receta=receta,
                total_cantidad=Decimal(qty), total_venta=Decimal(monto), total_venta_neta=Decimal(monto),
            )

        venta(proyectada, "Bollo Lotus", "500", 10)
        venta(temporada, "Rosca de Reyes", "800", 4)      # receta sin proyección
        venta(None, "Producto Nuevo SV", "300", 6)         # sin receta ligada

        resultado = comparativo_ventas_unidades(date(2026, 5, 1), hoy=date(2026, 5, 31))
        fuera = resultado["fuera_proyeccion"]
        nombres = [f["producto"] for f in fuera["top"]]
        self.assertIn("Rosca de Reyes", nombres)
        self.assertIn("Producto Nuevo SV", nombres)
        self.assertNotIn("Bollo Lotus", nombres)
        self.assertEqual(fuera["total_venta"], Decimal("1100"))
        self.assertEqual(fuera["total_unidades"], Decimal("10"))


class RenombradoPointTests(TestCase):
    """Los rubros de Ventas adoptan los nombres del catálogo Point."""

    @classmethod
    def setUpTestData(cls):
        cls.area = AreaPresupuesto.objects.create(nombre="Ventas", codigo="ventas")

    def _rubro_con_regla(self, concepto, filtros):
        rubro = RubroPresupuesto.objects.create(
            area=self.area, concepto=concepto, tipo=RubroPresupuesto.TIPO_INGRESO
        )
        ReglaFuenteRubro.objects.create(
            rubro=rubro, tipo_fuente=ReglaFuenteRubro.FUENTE_VENTA_POS, filtros=filtros
        )
        return rubro

    def test_renombra_a_producto_categoria_y_conserva_nombre_excel(self):
        uno = self._rubro_con_regla(
            "BOLLO · CHOCOLATE",
            {"productos_pos": ["Bollo Chocolate", "Bollo Chocolate SV"], "campo_monto": "total_venta"},
        )
        cat = self._rubro_con_regla(
            "BEBIDAS/OTROS · TE", {"categoria_pos": "TE", "campo_monto": "total_venta"}
        )
        pendiente = self._rubro_con_regla("CHEESECAKE · FRESA R", {"campo_monto": "total_venta"})

        call_command("renombrar_rubros_ventas_point", stdout=StringIO())

        uno.refresh_from_db(); cat.refresh_from_db(); pendiente.refresh_from_db()
        self.assertEqual(uno.concepto, "Bollo Chocolate")  # el más corto = producto base
        self.assertEqual(uno.metadata["nombre_excel"], "BOLLO · CHOCOLATE")
        self.assertEqual(cat.concepto, "TE")
        self.assertEqual(pendiente.concepto, "CHEESECAKE · FRESA R")  # sin asignación: intacto

        # Idempotente: segunda corrida no cambia nada ni pisa nombre_excel.
        call_command("renombrar_rubros_ventas_point", stdout=StringIO())
        uno.refresh_from_db()
        self.assertEqual(uno.concepto, "Bollo Chocolate")
        self.assertEqual(uno.metadata["nombre_excel"], "BOLLO · CHOCOLATE")

    def test_reimport_reconoce_nombre_excel_y_no_duplica(self):
        """Volver a importar el Excel actualiza el rubro renombrado, sin crear
        un duplicado con el nombre viejo."""
        import csv as csv_mod
        import tempfile

        from reportes.services_presupuesto_maestro import PresupuestoMaestroImportService

        rubro = self._rubro_con_regla(
            "BOLLO · CHOCOLATE", {"productos_pos": ["Bollo Chocolate"], "campo_monto": "total_venta"}
        )
        LineaPresupuestoMensual.objects.create(
            rubro=rubro, periodo=date(2026, 1, 1), monto_presupuesto=Decimal("10")
        )
        call_command("renombrar_rubros_ventas_point", stdout=StringIO())
        rubro.refresh_from_db()
        self.assertEqual(rubro.concepto, "Bollo Chocolate")

        campos = ["concepto", "tipo", "sucursal"] + [m for m, _ in
                  [("enero",1),("febrero",2),("marzo",3),("abril",4),("mayo",5),("junio",6),
                   ("julio",7),("agosto",8),("septiembre",9),("octubre",10),("noviembre",11),("diciembre",12)]]
        fila = {c: "0" for c in campos}
        fila.update({"concepto": "BOLLO · CHOCOLATE", "tipo": "INGRESO", "sucursal": "", "enero": "5000.00"})
        tmp = tempfile.NamedTemporaryFile("w", newline="", suffix=".csv", delete=False, encoding="utf-8")
        writer = csv_mod.DictWriter(tmp, fieldnames=campos)
        writer.writeheader(); writer.writerow(fila); tmp.close()

        PresupuestoMaestroImportService().import_file(
            archivo=tmp.name, area_code="ventas", version="ORIGINAL", year=2026
        )

        rubros = RubroPresupuesto.objects.filter(area=self.area)
        self.assertEqual(rubros.count(), 1)  # sin duplicado con nombre Excel
        rubro.refresh_from_db()
        self.assertEqual(rubro.concepto, "Bollo Chocolate")  # nombre Point se queda
        linea = LineaPresupuestoMensual.objects.get(rubro=rubro, periodo=date(2026, 1, 1))
        self.assertEqual(linea.monto_presupuesto, Decimal("5000.00"))  # presupuesto actualizado


class NormalizacionConceptosTests(TestCase):
    """Nomenclatura unificada: Primera mayúscula, acrónimos, acentos y typos."""

    def test_funcion_de_normalizacion(self):
        from reportes.management.commands.normalizar_conceptos_rubros import normalizar_concepto

        casos = {
            "SUELDO": "Sueldo",
            "Dias festivos": "Días festivos",
            "Imss": "IMSS",
            "Mantanimiento equipo/maquinaria": "Mantenimiento equipo/maquinaria",
            "Material de seguiridad e higiene": "Material de seguridad e higiene",
            "JUEGO DE LLANTAS PEGEOT": "Juego de llantas Peugeot",
            "SISTEMA DE REGRIFERACION MANAGER": "Sistema de refrigeración Manager",
            "Cuotas y suscriciones": "Cuotas y suscripciones",
            "Impuesto sobre Nómina": "Impuesto sobre nómina",
            "Mantenimiento eq. de computo": "Mantenimiento eq. de cómputo",
            # Palabras YA acentuadas encuentran su grafía (bug de primera versión).
            "CAPEX Guamúchil local": "CAPEX Guamúchil local",
            "Renta y agua Leyva": "Renta y agua Leyva",
        }
        for entrada, esperado in casos.items():
            self.assertEqual(normalizar_concepto(entrada), esperado, entrada)
        # Idempotencia: normalizar dos veces da lo mismo.
        for esperado in casos.values():
            self.assertEqual(normalizar_concepto(esperado), esperado)

    def test_comando_renombra_conserva_referencia_y_el_csv_sigue_cruzando(self):
        nomina = AreaPresupuesto.objects.create(nombre="Nómina", codigo="nomina")
        CategoriaGasto.objects.create(
            codigo="RENTA", nombre="Renta sucursal", capa_objetivo=CategoriaGasto.CAPA_EMPRESA
        )
        sueldo = RubroPresupuesto.objects.create(
            area=nomina, concepto="SUELDO", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        festivo = RubroPresupuesto.objects.create(
            area=nomina, concepto="FESTIVO", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        festivos = RubroPresupuesto.objects.create(
            area=nomina, concepto="FESTIVOS", tipo=RubroPresupuesto.TIPO_EGRESO
        )

        call_command("normalizar_conceptos_rubros", stdout=StringIO())

        sueldo.refresh_from_db(); festivo.refresh_from_db(); festivos.refresh_from_db()
        self.assertEqual(sueldo.concepto, "Sueldo")
        self.assertEqual(sueldo.metadata["nombre_excel"], "SUELDO")
        self.assertEqual(festivo.concepto, "Festivo")
        self.assertEqual(festivos.concepto, "Festivos")  # distintos, sin colisión

        # El seed sigue cruzando el CSV (matching insensible a caso/acentos):
        # la fila nomina,SUELDO,NOMINA aplica al rubro renombrado "Sueldo".
        call_command("seed_reglas_fuente_rubro", stdout=StringIO())
        self.assertTrue(
            ReglaFuenteRubro.objects.filter(
                rubro=sueldo, tipo_fuente=ReglaFuenteRubro.FUENTE_NOMINA
            ).exists()
        )


class CuraduriaRubrosTests(TestCase):
    """Fusión de duplicados y desactivación de rubros fantasma."""

    @classmethod
    def setUpTestData(cls):
        cls.logistica = AreaPresupuesto.objects.create(nombre="Logística", codigo="logistica")

    def _rubro(self, concepto, presupuesto="100", real=None, fuente=""):
        rubro = RubroPresupuesto.objects.create(
            area=self.logistica, concepto=concepto, tipo=RubroPresupuesto.TIPO_EGRESO
        )
        LineaPresupuestoMensual.objects.create(
            rubro=rubro, periodo=date(2026, 1, 1), monto_presupuesto=Decimal(presupuesto),
            monto_real=Decimal(real) if real else None, fuente_real=fuente,
        )
        return rubro

    def test_fusiona_suma_presupuesto_y_desactiva_origenes(self):
        destino_previo = self._rubro("Cheyenne", presupuesto="100")
        duplicado = self._rubro("Camioneta Cheyenne", presupuesto="50", real="30", fuente="MANUAL:jefa")

        call_command(
            "fusionar_rubros", area="logistica", destino="Chevrolet Cheyenne",
            origenes=["Cheyenne", "Camioneta Cheyenne"], stdout=StringIO(),
        )

        destino_previo.refresh_from_db(); duplicado.refresh_from_db()
        self.assertEqual(destino_previo.concepto, "Chevrolet Cheyenne")  # renombrado
        self.assertTrue(destino_previo.activo)
        self.assertFalse(duplicado.activo)
        linea = LineaPresupuestoMensual.objects.get(rubro=destino_previo, periodo=date(2026, 1, 1))
        self.assertEqual(linea.monto_presupuesto, Decimal("150"))  # sumado
        self.assertEqual(linea.monto_real, Decimal("30"))          # real conservado
        self.assertEqual(linea.fuente_real, "MANUAL:jefa")
        self.assertEqual(LineaPresupuestoMensual.objects.filter(rubro=duplicado).count(), 0)

    def test_fusion_no_pierde_reales_en_conflicto(self):
        self._rubro("Peugeot Manager", presupuesto="100", real="10", fuente="MANUAL:a")
        conflicto = self._rubro("PEUGEOT MANAGER", presupuesto="50", real="99", fuente="MANUAL:b")

        salida = StringIO()
        call_command(
            "fusionar_rubros", area="logistica", destino="Peugeot Manager",
            origenes=["PEUGEOT MANAGER"], stdout=salida,
        )
        self.assertIn("CONFLICTO", salida.getvalue())
        # La línea en conflicto NO se movió ni se borró.
        self.assertEqual(LineaPresupuestoMensual.objects.filter(rubro=conflicto).count(), 1)

    def test_desactivar_saca_del_tablero_y_protege_capturas(self):
        from django.contrib.auth import get_user_model

        fantasma = self._rubro("CHEESECAKE · FRESA R")
        con_captura = self._rubro("Con captura", real="5", fuente="MANUAL:x")

        call_command(
            "desactivar_rubros", area="logistica",
            conceptos=["CHEESECAKE · FRESA R", "Con captura"],
            motivo="no existe en Point", stdout=StringIO(),
        )
        fantasma.refresh_from_db(); con_captura.refresh_from_db()
        self.assertFalse(fantasma.activo)
        self.assertTrue(con_captura.activo)  # protegido sin --forzar

        # Y el tablero ya no lo incluye.
        superuser = get_user_model().objects.create_superuser("dg_cura", "c@test.mx", "x")
        self.client.force_login(superuser)
        response = self.client.get("/reportes/presupuesto-vs-real/?year=2026&month=1")
        conceptos = [r["concepto"] for r in response.context["detalle"]]
        self.assertNotIn("CHEESECAKE · FRESA R", conceptos)
        self.assertIn("Con captura", conceptos)


class FusionPorCuentaTests(TestCase):
    """Duplicados con el MISMO nombre se distinguen por cuenta contable."""

    def test_fusiona_mismo_nombre_por_cuenta(self):
        area = AreaPresupuesto.objects.create(nombre="Logística", codigo="logistica")
        con_cuenta = RubroPresupuesto.objects.create(
            area=area, concepto="Peugeot Partner", codigo_cuenta="1057-0003-0000",
            tipo=RubroPresupuesto.TIPO_EGRESO,
        )
        sin_cuenta = RubroPresupuesto.objects.create(
            area=area, concepto="Peugeot Partner", codigo_cuenta="",
            tipo=RubroPresupuesto.TIPO_EGRESO,
        )
        LineaPresupuestoMensual.objects.create(
            rubro=con_cuenta, periodo=date(2026, 1, 1), monto_presupuesto=Decimal("100")
        )
        LineaPresupuestoMensual.objects.create(
            rubro=sin_cuenta, periodo=date(2026, 1, 1), monto_presupuesto=Decimal("40")
        )

        call_command(
            "fusionar_rubros", area="logistica", destino="Peugeot Partner",
            origenes=["Peugeot Partner"], destino_cuenta="1057-0003-0000",
            origen_cuenta="", stdout=StringIO(),
        )

        con_cuenta.refresh_from_db(); sin_cuenta.refresh_from_db()
        self.assertTrue(con_cuenta.activo)
        self.assertFalse(sin_cuenta.activo)
        linea = LineaPresupuestoMensual.objects.get(rubro=con_cuenta)
        self.assertEqual(linea.monto_presupuesto, Decimal("140"))


class BonosYConsumoTests(TestCase):
    """PR4: bonos de producción/ventas y consumo de materia prima."""

    @classmethod
    def setUpTestData(cls):
        from bonos_produccion.models import BonoProduccionEmpleado, ConfigBonoPeriodo
        from bonos_ventas.models import BonoVentasEmpleado, ConfigBonoVentasPeriodo

        cls.periodo = date(2026, 6, 1)
        cls.sucursal = Sucursal.objects.create(codigo="BNS01", nombre="Centro bonos")
        cls.area = AreaPresupuesto.objects.create(nombre="Nómina", codigo="nomina")

        empleado = Empleado.objects.create(
            codigo="EMP-BONO-1", nombre="Prod uno", departamento=Empleado.DEP_PRODUCCION,
            sucursal_ref=cls.sucursal,
        )
        periodo_prod = ConfigBonoPeriodo.objects.create(mes=6, anio=2026)
        BonoProduccionEmpleado.objects.create(
            periodo=periodo_prod, empleado=empleado,
            monto_asistencia=Decimal("100"), monto_puntualidad=Decimal("50"),
            total_a_pagar=Decimal("400"),
        )
        periodo_ven = ConfigBonoVentasPeriodo.objects.create(mes=6, anio=2026)
        BonoVentasEmpleado.objects.create(
            periodo=periodo_ven, empleado=empleado, sucursal=cls.sucursal,
            monto_asistencia=Decimal("30"), monto_puntualidad=Decimal("20"),
            total_a_pagar=Decimal("250"),
        )

    def _linea(self, concepto, reglas):
        rubro = RubroPresupuesto.objects.create(
            area=self.area, concepto=concepto, tipo=RubroPresupuesto.TIPO_EGRESO
        )
        for kwargs in reglas:
            ReglaFuenteRubro.objects.create(rubro=rubro, **kwargs)
        return LineaPresupuestoMensual.objects.create(
            rubro=rubro, periodo=self.periodo, monto_presupuesto=Decimal("1")
        )

    def test_bonos_resultados_resta_componentes_sin_doble_conteo(self):
        F = ReglaFuenteRubro
        asistencia = self._linea("Bonos asistencia", [
            {"tipo_fuente": F.FUENTE_BONO_PRODUCCION, "filtros": {"campo_monto": "monto_asistencia"}},
            {"tipo_fuente": F.FUENTE_BONO_VENTAS, "filtros": {"campo_monto": "monto_asistencia"}},
        ])
        resultados = self._linea("Bonos por resultados", [
            {"tipo_fuente": F.FUENTE_BONO_PRODUCCION, "filtros": {"campo_monto": "total_a_pagar"}},
            {"tipo_fuente": F.FUENTE_BONO_PRODUCCION, "filtros": {"campo_monto": "monto_asistencia"}, "signo": -1},
            {"tipo_fuente": F.FUENTE_BONO_PRODUCCION, "filtros": {"campo_monto": "monto_puntualidad"}, "signo": -1},
            {"tipo_fuente": F.FUENTE_BONO_VENTAS, "filtros": {"campo_monto": "total_a_pagar"}},
            {"tipo_fuente": F.FUENTE_BONO_VENTAS, "filtros": {"campo_monto": "monto_asistencia"}, "signo": -1},
            {"tipo_fuente": F.FUENTE_BONO_VENTAS, "filtros": {"campo_monto": "monto_puntualidad"}, "signo": -1},
        ])

        PresupuestoRealConsolidacionService().consolidar(periodo=self.periodo)

        asistencia.refresh_from_db(); resultados.refresh_from_db()
        self.assertEqual(asistencia.monto_real, Decimal("130.00"))  # 100 prod + 30 ventas
        # resultados = (400-100-50) + (250-30-20) = 250 + 200 = 450
        self.assertEqual(resultados.monto_real, Decimal("450.00"))
        # Suma total de los rubros = total de ambos módulos (400+250) menos puntualidad aún sin rubro aquí
        # (asistencia 130 + resultados 450 + puntualidad 70 = 650 = 400+250) ✓ sin doble conteo.

    def test_consumo_mp_por_insumo(self):
        from inventario.models import ConsumoInsumoMensual
        from maestros.models import Insumo

        harina = Insumo.objects.create(nombre="Harina Espiga")
        ConsumoInsumoMensual.objects.create(
            insumo=harina, periodo=self.periodo, costo_real=Decimal("12345.67")
        )
        linea = self._linea("Harina espiga", [
            {"tipo_fuente": ReglaFuenteRubro.FUENTE_CONSUMO_MP, "filtros": {"insumo_id": harina.id}},
        ])

        PresupuestoRealConsolidacionService().consolidar(periodo=self.periodo)

        linea.refresh_from_db()
        self.assertEqual(linea.monto_real, Decimal("12345.67"))
        self.assertEqual(linea.fuente_real, "AUTO:CONSUMO_MP")

    def test_seed_asigna_insumos_de_produccion(self):
        from maestros.models import Insumo

        CategoriaGasto.objects.create(
            codigo="RENTA", nombre="Renta sucursal", capa_objetivo=CategoriaGasto.CAPA_EMPRESA
        )
        produccion = AreaPresupuesto.objects.create(nombre="Producción", codigo="produccion")
        azucar = Insumo.objects.create(nombre="Azúcar Estándar")
        Insumo.objects.create(nombre="Azúcar Glass")
        rubro = RubroPresupuesto.objects.create(
            area=produccion, concepto="Azucar estandar", tipo=RubroPresupuesto.TIPO_COSTO
        )
        LineaPresupuestoMensual.objects.create(
            rubro=rubro, periodo=self.periodo, monto_presupuesto=Decimal("1")
        )

        call_command("seed_reglas_fuente_rubro", stdout=StringIO())

        regla = ReglaFuenteRubro.objects.get(rubro=rubro, tipo_fuente=ReglaFuenteRubro.FUENTE_CONSUMO_MP)
        self.assertEqual(regla.filtros["insumo_id"], azucar.id)


class NavegacionCapturaTests(TestCase):
    """La captura aparece en el menú de responsables sin necesitar la liga."""

    def test_nav_muestra_captura_a_responsable_y_tablero_a_direccion(self):
        from django.contrib.auth import get_user_model

        from core.navigation import build_nav_groups
        from reportes.models import AreaPresupuestoResponsable

        User = get_user_model()
        area = AreaPresupuesto.objects.create(nombre="Logística", codigo="logistica")

        jefa = User.objects.create_user("jefa_nav", "jn@test.mx", "x")
        AreaPresupuestoResponsable.objects.create(area=area, usuario=jefa)
        etiquetas_jefa = [i["label"] for g in build_nav_groups(jefa, "/") for i in g["items"]]
        self.assertIn("Captura de presupuesto", etiquetas_jefa)
        self.assertNotIn("Presupuesto vs Real", etiquetas_jefa)

        ajeno = User.objects.create_user("ajeno_nav", "an@test.mx", "x")
        etiquetas_ajeno = [i["label"] for g in build_nav_groups(ajeno, "/") for i in g["items"]]
        self.assertNotIn("Captura de presupuesto", etiquetas_ajeno)

        dg = User.objects.create_superuser("dg_nav", "dn@test.mx", "x")
        etiquetas_dg = [i["label"] for g in build_nav_groups(dg, "/") for i in g["items"]]
        self.assertIn("Presupuesto vs Real", etiquetas_dg)
        self.assertIn("Captura de presupuesto", etiquetas_dg)


class ImportRealGastosExcelTests(TestCase):
    """El import de gastos captura la columna REAL con las protecciones."""

    def _xlsx_gastos(self, real_enero="1234.56"):
        import tempfile

        from openpyxl import Workbook

        wb = Workbook()
        hoja = wb.active
        hoja.title = "LOGISTICA"
        hoja.append(["CUENTA", "DESCRIPCION", "ENERO", "", "", "FEBRERO", "", ""])
        hoja.append(["", "", "PRESUPUESTADO", "REAL", "VARIACION", "PRESUPUESTADO", "REAL", "VARIACION"])
        hoja.append(["", "Imss", 5000, real_enero, "", 5000, "", ""])
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        return tmp.name

    def test_real_de_excel_entra_como_legado_y_respeta_manual(self):
        from reportes.services_presupuesto_maestro import PresupuestoMaestroImportService

        service = PresupuestoMaestroImportService()
        service.import_file(
            archivo=self._xlsx_gastos(), area_code="logistica", version="ORIGINAL", year=2026
        )
        linea = LineaPresupuestoMensual.objects.get(
            periodo=date(2026, 1, 1), rubro__concepto="Imss", rubro__area__codigo="logistica"
        )
        self.assertEqual(linea.monto_real, Decimal("1234.56"))
        self.assertEqual(linea.fuente_real, "AUTO:LEGADO")
        # Febrero no traía REAL: queda pendiente, no cero.
        febrero = LineaPresupuestoMensual.objects.get(
            periodo=date(2026, 2, 1), rubro__concepto="Imss", rubro__area__codigo="logistica"
        )
        self.assertIsNone(febrero.monto_real)

        # Captura manual posterior NO se pisa al re-importar.
        linea.fuente_real = "MANUAL:paula.lugo"
        linea.monto_real = Decimal("999")
        linea.save()
        service.import_file(
            archivo=self._xlsx_gastos(real_enero="5555"), area_code="logistica",
            version="ORIGINAL", year=2026,
        )
        linea.refresh_from_db()
        self.assertEqual(linea.monto_real, Decimal("999"))


class CedulaImssTests(TestCase):
    """Import de cédulas SIPARE: parseo por etiquetas, cruce NSS y reparto."""

    @classmethod
    def setUpTestData(cls):
        cls.sucursal = Sucursal.objects.create(codigo="CED01", nombre="Centro cédula")
        cls.gv = AreaPresupuesto.objects.create(nombre="Gastos de Venta", codigo="gastos-venta")
        cls.adm = AreaPresupuesto.objects.create(nombre="Administración", codigo="administracion")
        cls.nom = AreaPresupuesto.objects.create(nombre="Nómina", codigo="nomina")
        for area, sucursal in [(cls.gv, cls.sucursal), (cls.adm, None), (cls.nom, None)]:
            for concepto in ["IMSS", "Infonavit"]:
                RubroPresupuesto.objects.create(
                    area=area, concepto=concepto, sucursal=sucursal,
                    tipo=RubroPresupuesto.TIPO_EGRESO,
                )
        Empleado.objects.create(
            codigo="CED-1", nombre="Vendedora uno", nss="23-91-73-3507-9",
            departamento=Empleado.DEP_VENTAS, sucursal_ref=cls.sucursal,
        )
        Empleado.objects.create(
            codigo="CED-2", nombre="Admin uno", nss="23988051852",
            departamento=Empleado.DEP_ADMINISTRACION,
        )

    def _filas_mensual(self):
        return [
            [""] * 22,
            ["SISTEMA ÚNICO DE AUTODETERMINACIÓN"] + [""] * 21,
            [""] * 22,
            [""] * 22,
            [""] * 22,
            [""] * 22,
            ["Fecha de Proceso", "2026-05-15"] + [""] * 20,
            ["Período de Proceso: Abril-2026"] + [""] * 21,
            [""] * 22,
            ["Registro Patronal: ", "E52-40157-10-0"] + [""] * 20,
            *[[""] * 22 for _ in range(11)],
            ["Clave", "", "Fecha", "Días", "SDI", "Lic.", "Inc.", "Aus.", "C.F.", "Exc.Pat.",
             "Exc. Obr.", "P.D. Pat.", "P.D. Obr.", "G.M.P. Pat.", "G.M.P. Obr.", "R.T.",
             "I.V. Pat.", "I.V. Obr", "G.P.S.", "Patronal", "Obrera", "SubTotal"],
            ["23-91-73-3507-9", "", "", "", "", "ACOSTA FLORES MARIA"] + [""] * 16,
            [""] * 22,
            ["", "", "", 30, 331.44, 0, 3, 0, 646.14, 0, 0, 62.64, 22.37, 93.96, 33.56,
             44.74, 156.61, 55.93, 89.49, 1093.58, 111.86, 1205.44],
            ["23-98-80-5185-2", "", "", "", "", "PEREZ ADMIN JUAN"] + [""] * 16,
            [""] * 22,
            ["", "", "", 30, 331.01, 0, 0, 0, 717.94, 0, 0, 69.51, 24.83, 104.27, 37.24,
             49.65, 173.78, 62.06, 99.3, 1214.45, 124.13, 1338.58],
            ["99-99-99-9999-9", "", "", "", "", "SIN CRUCE FULANO"] + [""] * 16,
            [""] * 22,
            ["", "", "", 30, 300, 0, 0, 0, 500, 0, 0, 50, 20, 80, 30, 40, 150, 50, 80, 930, 100, 1030],
        ]

    def test_mensual_cruza_reparte_y_reporta_sin_nss(self):
        from reportes.services_cedula_imss import aplicar_cedula, parsear_cedula

        parseada = parsear_cedula(self._filas_mensual())
        self.assertEqual(parseada.tipo, "MENSUAL")
        self.assertEqual(parseada.periodo, date(2026, 4, 1))
        self.assertEqual(len(parseada.trabajadores), 3)

        resumen = aplicar_cedula(parseada)
        self.assertEqual(resumen.empleados_cruzados, 2)
        self.assertEqual(len(resumen.nss_sin_cruce), 1)

        imss_suc = LineaPresupuestoMensual.objects.get(
            rubro__area=self.gv, rubro__concepto="IMSS", rubro__sucursal=self.sucursal,
            periodo=date(2026, 4, 1),
        )
        self.assertEqual(imss_suc.monto_real, Decimal("1093.58"))  # patronal vendedora
        self.assertEqual(imss_suc.fuente_real, "AUTO:SIPARE")
        imss_adm = LineaPresupuestoMensual.objects.get(
            rubro__area=self.adm, rubro__concepto="IMSS", periodo=date(2026, 4, 1),
        )
        self.assertEqual(imss_adm.monto_real, Decimal("1214.45"))
        imss_nom = LineaPresupuestoMensual.objects.get(
            rubro__area=self.nom, rubro__concepto="IMSS", periodo=date(2026, 4, 1),
        )
        # El control (Nómina) lleva el total ÍNTEGRO de la cédula, incluido
        # el NSS sin cruce ($930) — el dinero nunca se esconde.
        self.assertEqual(imss_nom.monto_real, Decimal("3238.03"))

    def test_bimestral_parte_mitad_y_mitad_y_respeta_manual(self):
        from reportes.services_cedula_imss import aplicar_cedula, parsear_cedula

        filas = self._filas_mensual()
        filas[7] = ["Bimestre de Proceso: Abril-2026"] + [""] * 21
        filas[21] = ["Clave", "Fecha", "Días", "SDI", "Lic.", "Inc.", "Aus.", "Retiro",
                     "Patronal", "Obrera", "Suma", "Aportación Pa", "% o $ o FD", "*",
                     "Suma", "Créd. Vivienda", "Tipo"] + [""] * 5
        filas[24] = ["", "", 61, 331.44, 0, 3, 0, 404.36, 1158.41, 216.26, 1779.03,
                     1010.89, "", 1404.24, "", 2823.48, ""] + [""] * 5
        filas[27] = ["", "", 61, 331.01, 0, 0, 0, 403.83, 1216.75, 227.16, 1847.74,
                     1009.58, "", "", "", 0, ""] + [""] * 5
        filas[30] = ["", "", 61, 300, 0, 0, 0, 400, 1100, 200, 1700, 1000, "", "", "", 0, ""] + [""] * 5

        # Captura manual previa en marzo (mes 1 del bimestre): NO debe pisarse.
        infonavit_adm = RubroPresupuesto.objects.get(area=self.adm, concepto="Infonavit")
        LineaPresupuestoMensual.objects.create(
            rubro=infonavit_adm, periodo=date(2026, 3, 1), monto_presupuesto=Decimal("0"),
            monto_real=Decimal("777"), fuente_real="MANUAL:paula.lugo",
        )

        parseada = parsear_cedula(filas)
        self.assertEqual(parseada.tipo, "BIMESTRAL")
        self.assertEqual([m.isoformat() for m in parseada.meses], ["2026-03-01", "2026-04-01"])

        resumen = aplicar_cedula(parseada)
        # Vendedora: (404.36 + 1158.41 + 1010.89) = 2573.66 → 1286.83 por mes.
        marzo = LineaPresupuestoMensual.objects.get(
            rubro__area=self.gv, rubro__concepto="Infonavit", periodo=date(2026, 3, 1),
        )
        abril = LineaPresupuestoMensual.objects.get(
            rubro__area=self.gv, rubro__concepto="Infonavit", periodo=date(2026, 4, 1),
        )
        self.assertEqual(marzo.monto_real, Decimal("1286.83"))
        self.assertEqual(abril.monto_real, Decimal("1286.83"))
        # La captura manual de marzo en administración quedó intacta.
        linea_manual = LineaPresupuestoMensual.objects.get(
            rubro=infonavit_adm, periodo=date(2026, 3, 1),
        )
        self.assertEqual(linea_manual.monto_real, Decimal("777"))
        self.assertEqual(resumen.protegidas_manual, 1)


class CedulaImssEndurecidaTests(TestCase):
    """Fixes de la revisión adversarial del import de cédulas."""

    @classmethod
    def setUpTestData(cls):
        cls.adm = AreaPresupuesto.objects.create(nombre="Administración", codigo="administracion")
        cls.nom = AreaPresupuesto.objects.create(nombre="Nómina", codigo="nomina")
        for area in [cls.adm, cls.nom]:
            for concepto in ["IMSS", "Infonavit"]:
                RubroPresupuesto.objects.create(
                    area=area, concepto=concepto, tipo=RubroPresupuesto.TIPO_EGRESO
                )

    def _parseada(self, tipo="MENSUAL", periodo=None, trabajadores=None):
        from reportes.services_cedula_imss import CedulaParseada, TrabajadorCedula

        return CedulaParseada(
            tipo=tipo, periodo=periodo or date(2026, 4, 1), registro_patronal="E52",
            trabajadores=[TrabajadorCedula(*t) for t in (trabajadores or [])],
        )

    def test_area_con_mezcla_de_sucursales_no_truena(self):
        """Ventas con sucursal + sin sucursal en la misma corrida (bug de prod)."""
        from reportes.services_cedula_imss import aplicar_cedula

        gv = AreaPresupuesto.objects.create(nombre="Gastos de Venta", codigo="gastos-venta")
        sucursal = Sucursal.objects.create(codigo="MIX01", nombre="Mixta")
        RubroPresupuesto.objects.create(area=gv, concepto="IMSS", tipo=RubroPresupuesto.TIPO_EGRESO)
        RubroPresupuesto.objects.create(
            area=gv, concepto="IMSS", sucursal=sucursal, tipo=RubroPresupuesto.TIPO_EGRESO
        )
        Empleado.objects.create(
            codigo="MIX-1", nombre="Con sucursal", nss="77777777777",
            departamento=Empleado.DEP_VENTAS, sucursal_ref=sucursal,
        )
        Empleado.objects.create(
            codigo="MIX-2", nombre="Sin sucursal", nss="88888888888",
            departamento=Empleado.DEP_VENTAS,
        )
        resumen = aplicar_cedula(self._parseada(trabajadores=[
            ("77777777777", "A", Decimal("10")), ("88888888888", "B", Decimal("20")),
        ]))
        self.assertEqual(resumen.empleados_cruzados, 2)

    def test_nss_duplicado_en_rrhh_aborta(self):
        from reportes.services_cedula_imss import aplicar_cedula

        for codigo in ["DUP-1", "DUP-2"]:
            Empleado.objects.create(
                codigo=codigo, nombre=f"Empleado {codigo}", nss="11-11-11-1111-1",
                departamento=Empleado.DEP_ADMINISTRACION,
            )
        parseada = self._parseada(trabajadores=[("11111111111", "X", Decimal("100"))])
        with self.assertRaises(ValueError):
            aplicar_cedula(parseada)

    def test_total_control_incluye_sin_cruce_y_avisa(self):
        from reportes.services_cedula_imss import aplicar_cedula

        Empleado.objects.create(
            codigo="OK-1", nombre="Con cruce", nss="22222222222",
            departamento=Empleado.DEP_ADMINISTRACION,
        )
        parseada = self._parseada(trabajadores=[
            ("22222222222", "Con cruce", Decimal("100.00")),
            ("33333333333", "Sin cruce", Decimal("50.00")),
        ])
        resumen = aplicar_cedula(parseada)
        self.assertEqual(resumen.total_patronal, Decimal("150.00"))
        control = LineaPresupuestoMensual.objects.get(
            rubro__area=self.nom, rubro__concepto="IMSS", periodo=date(2026, 4, 1)
        )
        self.assertEqual(control.monto_real, Decimal("150.00"))  # íntegro
        adm = LineaPresupuestoMensual.objects.get(
            rubro__area=self.adm, rubro__concepto="IMSS", periodo=date(2026, 4, 1)
        )
        self.assertEqual(adm.monto_real, Decimal("100.00"))  # solo asignado
        self.assertTrue(any("SIN asignar" in a for a in resumen.avisos))

    def test_no_pisa_otra_fuente_auto_pero_si_legado(self):
        from reportes.services_cedula_imss import aplicar_cedula

        Empleado.objects.create(
            codigo="OK-2", nombre="Admin", nss="44444444444",
            departamento=Empleado.DEP_ADMINISTRACION,
        )
        rubro_adm = RubroPresupuesto.objects.get(area=self.adm, concepto="IMSS")
        otra_fuente = LineaPresupuestoMensual.objects.create(
            rubro=rubro_adm, periodo=date(2026, 4, 1), monto_presupuesto=Decimal("0"),
            monto_real=Decimal("999"), fuente_real="AUTO:GASTO_OPERATIVO",
        )
        resumen = aplicar_cedula(self._parseada(trabajadores=[("44444444444", "A", Decimal("80"))]))
        otra_fuente.refresh_from_db()
        self.assertEqual(otra_fuente.monto_real, Decimal("999"))  # conflicto, no pisado
        self.assertTrue(any("conflicto de fuentes" in a for a in resumen.avisos))

        # AUTO:LEGADO (Excel tecleado) SÍ cede ante la cédula oficial.
        otra_fuente.fuente_real = "AUTO:LEGADO"
        otra_fuente.save()
        aplicar_cedula(self._parseada(trabajadores=[("44444444444", "A", Decimal("80"))]))
        otra_fuente.refresh_from_db()
        self.assertEqual(otra_fuente.monto_real, Decimal("80.00"))
        self.assertEqual(otra_fuente.fuente_real, "AUTO:SIPARE")

    def test_bimestre_impar_rechazado_y_centavo_conservado(self):
        from reportes.services_cedula_imss import aplicar_cedula

        with self.assertRaises(ValueError):
            aplicar_cedula(self._parseada(
                tipo="BIMESTRAL", periodo=date(2026, 3, 1),
                trabajadores=[("55555555555", "X", Decimal("100"))],
            ))

        Empleado.objects.create(
            codigo="OK-3", nombre="Admin b", nss="66666666666",
            departamento=Empleado.DEP_ADMINISTRACION,
        )
        # Total con centavo impar: 100.01 → 50.01 + 50.00 (suma exacta).
        aplicar_cedula(self._parseada(
            tipo="BIMESTRAL", periodo=date(2026, 4, 1),
            trabajadores=[("66666666666", "B", Decimal("100.01"))],
        ))
        marzo = LineaPresupuestoMensual.objects.get(
            rubro__area=self.adm, rubro__concepto="Infonavit", periodo=date(2026, 3, 1)
        )
        abril = LineaPresupuestoMensual.objects.get(
            rubro__area=self.adm, rubro__concepto="Infonavit", periodo=date(2026, 4, 1)
        )
        self.assertEqual(marzo.monto_real + abril.monto_real, Decimal("100.01"))


class PantallaCedulaImssTests(TestCase):
    """Pantalla de subida de cédulas: RBAC, preview y aplicación."""

    URL = "/reportes/presupuesto-real/cedula-imss/"

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model

        User = get_user_model()
        cls.dg = User.objects.create_superuser("dg_ced", "dgc@test.mx", "x")
        cls.paula = User.objects.create_user("paula_ced", "p@test.mx", "x")
        cls.ajeno = User.objects.create_user("ajeno_ced", "a@test.mx", "x")

        from reportes.models import AreaPresupuestoResponsable

        nomina = AreaPresupuesto.objects.create(nombre="Nómina", codigo="nomina")
        adm = AreaPresupuesto.objects.create(nombre="Administración", codigo="administracion")
        for area in [nomina, adm]:
            RubroPresupuesto.objects.create(area=area, concepto="IMSS", tipo=RubroPresupuesto.TIPO_EGRESO)
        AreaPresupuestoResponsable.objects.create(area=nomina, usuario=cls.paula)
        Empleado.objects.create(
            codigo="PC-1", nombre="Admin cédula", nss="12121212121",
            departamento=Empleado.DEP_ADMINISTRACION,
        )

    def _subir(self, usuario, previsualizar=True):
        from unittest.mock import patch

        from django.core.files.uploadedfile import SimpleUploadedFile

        filas = [
            ["Período de Proceso: Abril-2026"],
            ["Clave", "", "", "Patronal", "Obrera"],
            ["12-12-12-1212-1", "TRABAJADORA UNO"],
            ["", "", "", 500.0, 100.0],
        ]
        self.client.force_login(usuario)
        datos = {"cedula": SimpleUploadedFile("cedula.xls", b"fake")}
        if previsualizar:
            datos["previsualizar"] = "1"
        with patch("reportes.services_cedula_imss.cargar_filas_xls", return_value=filas):
            return self.client.post(self.URL, datos)

    def test_rbac(self):
        self.client.force_login(self.ajeno)
        self.assertEqual(self.client.get(self.URL).status_code, 403)
        self.client.force_login(self.paula)
        self.assertEqual(self.client.get(self.URL).status_code, 200)  # responsable de nómina

    def test_preview_no_escribe_y_aplicar_si(self):
        r = self._subir(self.paula, previsualizar=True)
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.context["resumen"].lineas_actualizadas, 2)
        self.assertFalse(
            LineaPresupuestoMensual.objects.filter(fuente_real="AUTO:SIPARE").exists()
        )

        r = self._subir(self.paula, previsualizar=False)
        self.assertEqual(r.status_code, 200)
        linea = LineaPresupuestoMensual.objects.get(
            rubro__area__codigo="administracion", rubro__concepto="IMSS", periodo=date(2026, 4, 1)
        )
        self.assertEqual(linea.monto_real, Decimal("500.00"))
        self.assertEqual(linea.fuente_real, "AUTO:SIPARE")

    def test_extension_invalida_rechazada(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        self.client.force_login(self.dg)
        r = self.client.post(self.URL, {"cedula": SimpleUploadedFile("cedula.pdf", b"x")})
        self.assertIn(".xls", r.context["error"])


class AccesoAnonimoTests(TestCase):
    """Usuarios no autenticados van al login, nunca a un error 500."""

    def test_pantallas_redirigen_a_login(self):
        for url in [
            "/reportes/presupuesto-vs-real/",
            "/reportes/presupuesto-real/captura/",
            "/reportes/presupuesto-real/cedula-imss/",
        ]:
            respuesta = self.client.get(url)
            self.assertEqual(respuesta.status_code, 302, url)
            self.assertIn("login", respuesta["Location"], url)


class OverrideDirigidoTests(TestCase):
    """Dirección sobrescribe automáticos con motivo y puede liberar capturas."""

    URL_GUARDAR = "/reportes/presupuesto-real/captura/guardar/"
    URL_LIBERAR = "/reportes/presupuesto-real/liberar/"

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model

        from reportes.models import AreaPresupuestoResponsable

        User = get_user_model()
        cls.dg = User.objects.create_superuser("dg_ovr", "d@test.mx", "x")
        cls.jefa = User.objects.create_user("jefa_ovr", "j@test.mx", "x")

        area = AreaPresupuesto.objects.create(nombre="Logística", codigo="logistica")
        AreaPresupuestoResponsable.objects.create(area=area, usuario=cls.jefa)
        rubro_auto = RubroPresupuesto.objects.create(
            area=area, concepto="Sueldo logística", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        ReglaFuenteRubro.objects.create(
            rubro=rubro_auto,
            tipo_fuente=ReglaFuenteRubro.FUENTE_NOMINA,
            filtros={"campo_monto": "salario_base", "departamento": "LOGISTICA"},
        )
        cls.linea_auto = LineaPresupuestoMensual.objects.create(
            rubro=rubro_auto, periodo=date(2026, 6, 1), monto_presupuesto=Decimal("100"),
            monto_real=Decimal("80"), fuente_real="AUTO:NOMINA",
        )

    def test_jefa_no_sobrescribe_automatico(self):
        self.client.force_login(self.jefa)
        r = self.client.post(
            self.URL_GUARDAR,
            {"linea_id": self.linea_auto.id, "monto": "50", "motivo": "intento"},
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(r.status_code, 409)

    def test_direccion_sobrescribe_con_motivo_y_libera(self):
        self.client.force_login(self.dg)
        # Sin motivo → rechazado.
        r = self.client.post(
            self.URL_GUARDAR,
            {"linea_id": self.linea_auto.id, "monto": "77"},
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(r.status_code, 400)
        # Con motivo → queda MANUAL:dg con el motivo en el historial.
        r = self.client.post(
            self.URL_GUARDAR,
            {"linea_id": self.linea_auto.id, "monto": "77", "motivo": "Ajuste por finiquito"},
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(r.status_code, 200)
        self.linea_auto.refresh_from_db()
        self.assertEqual(self.linea_auto.monto_real, Decimal("77.00"))
        self.assertEqual(self.linea_auto.fuente_real, "MANUAL:dg_ovr")
        self.assertEqual(self.linea_auto.metadata["capturas"][-1]["motivo"], "Ajuste por finiquito")
        # Y el motor la respeta.
        PresupuestoRealConsolidacionService().consolidar(periodo=date(2026, 6, 1))
        self.linea_auto.refresh_from_db()
        self.assertEqual(self.linea_auto.monto_real, Decimal("77.00"))

        # Liberar: vuelve a pendiente y el motor la rellena de nuevo.
        r = self.client.post(
            self.URL_LIBERAR,
            {"linea_id": self.linea_auto.id},
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(r.status_code, 200)
        self.linea_auto.refresh_from_db()
        self.assertEqual(self.linea_auto.fuente_real, "")
        self.assertEqual(self.linea_auto.metadata["capturas"][-1]["accion"], "liberado_a_automatico")

    def test_jefa_no_puede_liberar(self):
        # Deja una captura manual primero (de la propia dirección).
        LineaPresupuestoMensual.objects.filter(pk=self.linea_auto.pk).update(
            fuente_real="MANUAL:dg_ovr"
        )
        self.client.force_login(self.jefa)
        r = self.client.post(
            self.URL_LIBERAR,
            {"linea_id": self.linea_auto.id},
            HTTP_ACCEPT="application/json",
        )
        self.assertEqual(r.status_code, 403)


class AreaResultadosTests(TestCase):
    """El bloque P&L de empresa sale de administración a un área de control."""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model

        cls.superuser = get_user_model().objects.create_superuser(
            "dg_resultados", "dgr@test.mx", "clave-test"
        )
        cls.periodo = date(2026, 3, 1)
        cls.admin_area = AreaPresupuesto.objects.create(
            nombre="Administración", codigo="administracion"
        )
        # El import del Excel tipó las ventas como EGRESO; el comando lo corrige.
        cls.rubro_pnl = RubroPresupuesto.objects.create(
            area=cls.admin_area, concepto="Venta postres", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        cls.rubro_gasto = RubroPresupuesto.objects.create(
            area=cls.admin_area, concepto="Sueldo", tipo=RubroPresupuesto.TIPO_EGRESO
        )

    def test_comando_mueve_solo_rubros_pnl(self):
        salida = StringIO()
        call_command("mover_rubros_resultados", stdout=salida)
        self.rubro_pnl.refresh_from_db()
        self.rubro_gasto.refresh_from_db()
        area = AreaPresupuesto.objects.get(codigo="resultados")
        self.assertEqual(self.rubro_pnl.area, area)
        self.assertEqual(self.rubro_pnl.tipo, RubroPresupuesto.TIPO_INGRESO)
        self.assertEqual(self.rubro_gasto.area, self.admin_area)
        # Idempotente: segunda corrida no truena ni duplica el área.
        call_command("mover_rubros_resultados", stdout=salida)
        self.assertEqual(AreaPresupuesto.objects.filter(codigo="resultados").count(), 1)

    def test_dry_run_no_mueve(self):
        salida = StringIO()
        call_command("mover_rubros_resultados", "--dry-run", stdout=salida)
        self.rubro_pnl.refresh_from_db()
        self.assertEqual(self.rubro_pnl.area, self.admin_area)
        self.assertFalse(AreaPresupuesto.objects.filter(codigo="resultados").exists())

    def test_kpi_global_excluye_area_resultados(self):
        call_command("mover_rubros_resultados", stdout=StringIO())
        LineaPresupuestoMensual.objects.create(
            rubro=self.rubro_pnl, periodo=self.periodo,
            monto_presupuesto=Decimal("4000000.00"), monto_real=Decimal("3500000.00"),
            fuente_real="AUTO:LEGADO",
        )
        LineaPresupuestoMensual.objects.create(
            rubro=self.rubro_gasto, periodo=self.periodo,
            monto_presupuesto=Decimal("100.00"), monto_real=Decimal("80.00"),
            fuente_real="AUTO:NOMINA",
        )
        self.client.force_login(self.superuser)
        response = self.client.get("/reportes/presupuesto-vs-real/?year=2026&month=3")
        kpis = response.context["kpis"]
        self.assertEqual(kpis["presupuesto"], Decimal("100.00"))
        self.assertEqual(kpis["real"], Decimal("80.00"))
        # Seleccionando el área de control sí se ve su propio total.
        response = self.client.get("/reportes/presupuesto-vs-real/?year=2026&month=3&area=resultados")
        self.assertEqual(response.context["kpis"]["presupuesto"], Decimal("4000000.00"))


class EstadoResultadosTests(TestCase):
    """P&L empresa completa: utilidades calculadas y MP sin doble conteo."""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model

        from maestros.models import Insumo

        cls.superuser = get_user_model().objects.create_superuser(
            "dg_pnl", "dgp@test.mx", "clave-test"
        )
        cls.periodo = date(2026, 2, 1)

        def rubro(area, concepto, tipo=RubroPresupuesto.TIPO_EGRESO):
            return RubroPresupuesto.objects.create(area=area, concepto=concepto, tipo=tipo)

        def linea(rubro_obj, ppto, real=None, periodo=None):
            return LineaPresupuestoMensual.objects.create(
                rubro=rubro_obj, periodo=periodo or cls.periodo,
                monto_presupuesto=Decimal(ppto),
                monto_real=Decimal(real) if real is not None else None,
                fuente_real="AUTO:LEGADO" if real is not None else "",
            )

        resultados = AreaPresupuesto.objects.create(nombre="Resultados", codigo="resultados")
        gastos = AreaPresupuesto.objects.create(nombre="Gastos de venta", codigo="gastos-venta")
        produccion = AreaPresupuesto.objects.create(nombre="Producción", codigo="produccion")
        capex = AreaPresupuesto.objects.create(nombre="CAPEX", codigo="capex")
        nomina = AreaPresupuesto.objects.create(nombre="Nómina", codigo="nomina")

        linea(rubro(resultados, "Venta postres", RubroPresupuesto.TIPO_INGRESO), "1000", "900")
        linea(rubro(resultados, "Costos insumos"), "400", "350")
        linea(rubro(gastos, "Renta"), "100", "90")
        rubro_mo = rubro(produccion, "Mano de obra")
        linea(rubro_mo, "50", "40")
        # Materia prima en producción: tiene regla CONSUMO_MP → NO debe sumar
        rubro_mp = rubro(produccion, "Queso crema")
        ReglaFuenteRubro.objects.create(
            rubro=rubro_mp, tipo_fuente=ReglaFuenteRubro.FUENTE_CONSUMO_MP,
            filtros={},
        )
        linea(rubro_mp, "400", "350")
        # Roll-up del Excel e insumo suelto sin regla: tampoco deben sumar
        linea(rubro(produccion, "Costo de producción"), "800", "700")
        Insumo.objects.create(nombre="Chocolate Blanco Turín")
        linea(rubro(produccion, "Chocolate blanco turin"), "60", "55")
        # Herramienta con nombre en catálogo: SÍ suma (no es materia prima)
        Insumo.objects.create(nombre="Mesa de trabajo", tipo_item=Insumo.TIPO_HERRAMIENTA)
        linea(rubro(produccion, "Mesa de trabajo"), "30", "20")
        linea(rubro(capex, "Horno nuevo"), "200", "100")
        linea(rubro(nomina, "Sueldos control"), "9999", "9999")
        # Marzo: solo presupuesto, sin real → columnas reales en blanco
        linea(rubro(gastos, "Renta marzo"), "100", periodo=date(2026, 3, 1))

    def _get(self):
        self.client.force_login(self.superuser)
        return self.client.get("/reportes/estado-resultados/?year=2026")

    def test_utilidades_y_exclusiones(self):
        filas = {f["label"]: f for f in self._get().context["filas"]}
        feb = 1  # índice del mes 2
        # Utilidad bruta = 1000-400 ppto, 900-350 real
        self.assertEqual(filas["Utilidad bruta"]["meses"][feb]["ppto"], Decimal("600"))
        self.assertEqual(filas["Utilidad bruta"]["meses"][feb]["real"], Decimal("550"))
        # Producción excluye MP con regla, roll-up e insumo suelto;
        # queda mano de obra (50/40) + herramienta (30/20)
        self.assertEqual(filas["Producción (sin materia prima)"]["meses"][feb]["ppto"], Decimal("80"))
        self.assertEqual(filas["Producción (sin materia prima)"]["meses"][feb]["real"], Decimal("60"))
        # Operativa = 600-180 ppto, 550-150 real
        self.assertEqual(filas["Utilidad operativa"]["meses"][feb]["ppto"], Decimal("420"))
        self.assertEqual(filas["Utilidad operativa"]["meses"][feb]["real"], Decimal("400"))
        # Resultado final = operativa - capex
        self.assertEqual(filas["Resultado final"]["meses"][feb]["ppto"], Decimal("220"))
        self.assertEqual(filas["Resultado final"]["meses"][feb]["real"], Decimal("300"))
        # Inversiones separadas: "Horno nuevo" no empieza con CAPEX → equipo
        self.assertEqual(filas["Compras de equipo"]["meses"][feb]["ppto"], Decimal("200"))
        self.assertEqual(filas["Inversión en proyectos (aperturas)"]["meses"][feb]["ppto"], Decimal("0"))

    def test_mes_sin_real_queda_en_blanco_y_nomina_fuera(self):
        filas = {f["label"]: f for f in self._get().context["filas"]}
        mar = 2
        self.assertIsNone(filas["Gastos de venta"]["meses"][mar]["real"])
        self.assertEqual(filas["Gastos de venta"]["meses"][mar]["ppto"], Decimal("100"))
        # Nómina (control) no aparece en ninguna fila
        anual_egresos = sum(f["anual_ppto"] for f in filas.values() if f["kind"] == "linea")
        self.assertLess(anual_egresos, Decimal("9999"))
        # Anual real de utilidad solo suma meses con ingresos reales
        self.assertEqual(filas["Resultado final"]["anual_real"], Decimal("300"))

    def test_requiere_login(self):
        response = self.client.get("/reportes/estado-resultados/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/login", response["Location"])


class HojaGeneralNoImportaTests(TestCase):
    """La hoja GENERAL (consolidado) no debe importarse como sucursal fantasma."""

    def _xlsx_con_general(self):
        import tempfile

        from openpyxl import Workbook

        wb = Workbook()
        hoja = wb.active
        hoja.title = "MATRIZ"
        hoja.append(["CUENTA", "DESCRIPCION", "ENERO", "", ""])
        hoja.append(["", "", "PRESUPUESTADO", "REAL", "VARIACION"])
        hoja.append(["", "Renta", 1000, 900, ""])
        general = wb.create_sheet("GENERAL")
        general.append(["CUENTA", "DESCRIPCION", "ENERO", "", ""])
        general.append(["", "", "PRESUPUESTADO", "REAL", "VARIACION"])
        general.append(["", "Renta", 1000, 900, ""])
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        return tmp.name

    def test_hoja_general_se_ignora(self):
        from core.models import Sucursal
        from reportes.services_presupuesto_maestro import PresupuestoMaestroImportService

        Sucursal.objects.create(nombre="Matriz", codigo="MATRIZ")
        service = PresupuestoMaestroImportService()
        service.import_file(
            archivo=self._xlsx_con_general(), area_code="gastos-venta",
            version="ORIGINAL", year=2026,
        )
        rubros = RubroPresupuesto.objects.filter(area__codigo="gastos-venta", concepto="Renta")
        self.assertEqual(rubros.count(), 1)
        self.assertIsNotNone(rubros.first().sucursal_id)

    def test_comando_desactiva_rubros_fantasma(self):
        area = AreaPresupuesto.objects.create(nombre="Gastos", codigo="gastos-venta")
        fantasma = RubroPresupuesto.objects.create(
            area=area, concepto="Renta", tipo=RubroPresupuesto.TIPO_EGRESO,
            metadata={"source": "PAQUETE_2026_REAL"},
        )
        otro_origen = RubroPresupuesto.objects.create(
            area=area, concepto="Rubro legítimo", tipo=RubroPresupuesto.TIPO_EGRESO,
            metadata={"source": "OTRO"},
        )
        salida = StringIO()
        call_command("desactivar_rubros_hoja_general", stdout=salida)
        fantasma.refresh_from_db()
        otro_origen.refresh_from_db()
        self.assertFalse(fantasma.activo)
        self.assertIn("desactivado_motivo", fantasma.metadata)
        self.assertTrue(otro_origen.activo)

    def test_dry_run_no_desactiva(self):
        area = AreaPresupuesto.objects.create(nombre="Gastos", codigo="gastos-venta")
        fantasma = RubroPresupuesto.objects.create(
            area=area, concepto="Renta", tipo=RubroPresupuesto.TIPO_EGRESO,
            metadata={"source": "PAQUETE_2026_REAL"},
        )
        call_command("desactivar_rubros_hoja_general", "--dry-run", stdout=StringIO())
        fantasma.refresh_from_db()
        self.assertTrue(fantasma.activo)


class TotalEmpresaReglasTests(TestCase):
    """Reglas total_empresa: los renglones del P&L leen fuentes vivas."""

    @classmethod
    def setUpTestData(cls):
        cls.periodo = date(2026, 6, 1)
        cls.area = AreaPresupuesto.objects.create(nombre="Resultados", codigo="resultados")
        cls.rubro_ventas = RubroPresupuesto.objects.create(
            area=cls.area, concepto="Venta postres", tipo=RubroPresupuesto.TIPO_INGRESO
        )
        cls.rubro_costos = RubroPresupuesto.objects.create(
            area=cls.area, concepto="Costos insumos/productos", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        LineaPresupuestoMensual.objects.create(
            rubro=cls.rubro_ventas, periodo=cls.periodo,
            monto_presupuesto=Decimal("100"), monto_real=Decimal("55"),
            fuente_real="AUTO:LEGADO",  # el Excel parcial que debe pisarse
        )
        LineaPresupuestoMensual.objects.create(
            rubro=cls.rubro_costos, periodo=cls.periodo, monto_presupuesto=Decimal("50"),
        )

    def test_venta_total_empresa_pisa_legado(self):
        sucursal = Sucursal.objects.create(nombre="Matriz", codigo="MATRIZ")
        branch = PointBranch.objects.create(external_id="b1", name="Matriz", erp_branch=sucursal)
        for categoria, producto, monto in (("PASTELES", "Pastel X", "30"), ("BOLLOS", "Bollo Y", "12")):
            PointSalesDailyProductFact.objects.create(
                branch=branch, sale_date=self.periodo, sucursal_nombre=branch.name,
                categoria=categoria, producto_nombre_historico=producto,
                total_venta=Decimal(monto), total_venta_neta=Decimal(monto),
            )
        ReglaFuenteRubro.objects.create(
            rubro=self.rubro_ventas, tipo_fuente=ReglaFuenteRubro.FUENTE_VENTA_POS,
            filtros={"total_empresa": True, "campo_monto": "total_venta"},
        )
        PresupuestoRealConsolidacionService().consolidar(periodo=self.periodo)
        linea = LineaPresupuestoMensual.objects.get(rubro=self.rubro_ventas, periodo=self.periodo)
        self.assertEqual(linea.monto_real, Decimal("42"))
        self.assertEqual(linea.fuente_real, "AUTO:VENTA_POS")

    def test_consumo_total_empresa(self):
        from inventario.models import ConsumoInsumoMensual
        from maestros.models import Insumo

        a = Insumo.objects.create(nombre="Harina P&L")
        b = Insumo.objects.create(nombre="Azúcar P&L")
        ConsumoInsumoMensual.objects.create(insumo=a, periodo=self.periodo, costo_real=Decimal("20"))
        ConsumoInsumoMensual.objects.create(insumo=b, periodo=self.periodo, costo_real=Decimal("15"))
        ReglaFuenteRubro.objects.create(
            rubro=self.rubro_costos, tipo_fuente=ReglaFuenteRubro.FUENTE_CONSUMO_MP,
            filtros={"total_empresa": True},
        )
        PresupuestoRealConsolidacionService().consolidar(periodo=self.periodo)
        linea = LineaPresupuestoMensual.objects.get(rubro=self.rubro_costos, periodo=self.periodo)
        self.assertEqual(linea.monto_real, Decimal("35"))
        self.assertEqual(linea.fuente_real, "AUTO:CONSUMO_MP")

    def test_limpiar_sin_asignacion_respeta_total_empresa(self):
        from reportes.services_presupuesto_real import limpiar_reales_sin_asignacion

        ReglaFuenteRubro.objects.create(
            rubro=self.rubro_ventas, tipo_fuente=ReglaFuenteRubro.FUENTE_VENTA_POS,
            filtros={"total_empresa": True},
        )
        limpiadas = limpiar_reales_sin_asignacion()
        self.assertEqual(limpiadas, 0)
        linea = LineaPresupuestoMensual.objects.get(rubro=self.rubro_ventas, periodo=self.periodo)
        self.assertEqual(linea.fuente_real, "AUTO:LEGADO")


class ConsumoDesdeFiltroTests(TestCase):
    """El filtro 'desde' de CONSUMO_MP protege los meses legados del Excel."""

    @classmethod
    def setUpTestData(cls):
        from inventario.models import ConsumoInsumoMensual
        from maestros.models import Insumo

        cls.area = AreaPresupuesto.objects.create(nombre="Resultados", codigo="resultados")
        cls.rubro = RubroPresupuesto.objects.create(
            area=cls.area, concepto="Costos insumos/productos", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        insumo = Insumo.objects.create(nombre="Harina desde-test")
        for mes, costo in ((3, "999"), (6, "500")):
            LineaPresupuestoMensual.objects.create(
                rubro=cls.rubro, periodo=date(2026, mes, 1),
                monto_presupuesto=Decimal("1000"),
                monto_real=Decimal("777"), fuente_real="AUTO:LEGADO",
            )
            ConsumoInsumoMensual.objects.create(
                insumo=insumo, periodo=date(2026, mes, 1), costo_real=Decimal(costo)
            )
        ReglaFuenteRubro.objects.create(
            rubro=cls.rubro, tipo_fuente=ReglaFuenteRubro.FUENTE_CONSUMO_MP,
            filtros={"total_empresa": True, "desde": "2026-06"},
        )

    def test_antes_de_desde_conserva_legado_y_despues_consolida(self):
        PresupuestoRealConsolidacionService().consolidar(periodo=date(2026, 3, 1))
        PresupuestoRealConsolidacionService().consolidar(periodo=date(2026, 6, 1))
        marzo = LineaPresupuestoMensual.objects.get(rubro=self.rubro, periodo=date(2026, 3, 1))
        junio = LineaPresupuestoMensual.objects.get(rubro=self.rubro, periodo=date(2026, 6, 1))
        self.assertEqual(marzo.monto_real, Decimal("777"))
        self.assertEqual(marzo.fuente_real, "AUTO:LEGADO")
        self.assertEqual(junio.monto_real, Decimal("500"))
        self.assertEqual(junio.fuente_real, "AUTO:CONSUMO_MP")

    def test_restaurar_costos_legado(self):
        linea = LineaPresupuestoMensual.objects.get(rubro=self.rubro, periodo=date(2026, 3, 1))
        linea.monto_real = Decimal("363918")
        linea.fuente_real = "AUTO:CONSUMO_MP"
        linea.save()
        salida = StringIO()
        call_command("restaurar_costos_pnl_legado", stdout=salida)
        linea.refresh_from_db()
        self.assertEqual(linea.monto_real, Decimal("1247660.66"))
        self.assertEqual(linea.fuente_real, "AUTO:LEGADO")
        # Idempotente
        call_command("restaurar_costos_pnl_legado", stdout=salida)
        self.assertIn("ya restaurado", salida.getvalue())


class CorregirPronosticoNombresPointTests(TestCase):
    """Reasignación de pronósticos a recetas Point vigentes (regla de dirección)."""

    @classmethod
    def setUpTestData(cls):
        from recetas.models import PronosticoVenta, Receta

        cls.sv = Receta.objects.create(nombre="3 Pecados Mini SV", hash_contenido="t-sv")
        cls.vigente = Receta.objects.create(nombre="Pastel 3 Pecados Mini", hash_contenido="t-mini")
        cls.sv_chico = Receta.objects.create(
            nombre="Pastel Fresas con Crema San Valentín Chico", hash_contenido="t-svc"
        )
        cls.chico = Receta.objects.create(
            nombre="Pastel de Fresas Con Crema Chico", hash_contenido="t-chico"
        )
        cls.flan = Receta.objects.create(nombre="Flan 3 Pecados Chico", hash_contenido="t-flan")
        PronosticoVenta.objects.create(receta=cls.sv, periodo="2026-05", cantidad=Decimal("194"), fuente="PRESUPUESTO_2026")
        PronosticoVenta.objects.create(receta=cls.sv_chico, periodo="2026-05", cantidad=Decimal("485"), fuente="PRESUPUESTO_2026")
        PronosticoVenta.objects.create(receta=cls.chico, periodo="2026-05", cantidad=Decimal("100"), fuente="PRESUPUESTO_2026")
        PronosticoVenta.objects.create(receta=cls.flan, periodo="2026-05", cantidad=Decimal("555"), fuente="PRESUPUESTO_2026")

    def test_reasigna_suma_y_elimina_internos(self):
        from recetas.models import PronosticoVenta

        call_command("corregir_pronostico_nombres_point", stdout=StringIO())
        # SV sin destino previo: se movió la fila entera
        self.assertEqual(PronosticoVenta.objects.filter(receta=self.sv).count(), 0)
        movido = PronosticoVenta.objects.get(receta=self.vigente, periodo="2026-05")
        self.assertEqual(movido.cantidad, Decimal("194"))
        # SV con destino previo: se sumaron cantidades
        sumado = PronosticoVenta.objects.get(receta=self.chico, periodo="2026-05")
        self.assertEqual(sumado.cantidad, Decimal("585"))
        self.assertEqual(PronosticoVenta.objects.filter(receta=self.sv_chico).count(), 0)
        # Insumo interno: eliminado del comparativo de ventas
        self.assertEqual(PronosticoVenta.objects.filter(receta=self.flan).count(), 0)
        # Idempotente
        call_command("corregir_pronostico_nombres_point", stdout=StringIO())
        self.assertEqual(PronosticoVenta.objects.get(receta=self.vigente, periodo="2026-05").cantidad, Decimal("194"))

    def test_dry_run_no_toca(self):
        from recetas.models import PronosticoVenta

        call_command("corregir_pronostico_nombres_point", "--dry-run", stdout=StringIO())
        self.assertEqual(PronosticoVenta.objects.filter(receta=self.sv).count(), 1)
        self.assertEqual(PronosticoVenta.objects.filter(receta=self.flan).count(), 1)


class ProductosDescontinuadosTests(TestCase):
    """Los productos descontinuados salen del pronóstico (dirección 2026-07-16)."""

    def test_elimina_descontinuados(self):
        from recetas.models import PronosticoVenta, Receta

        receta = Receta.objects.create(nombre="Brownie Rebanada", hash_contenido="t-br")
        PronosticoVenta.objects.create(receta=receta, periodo="2026-05", cantidad=Decimal("8"), fuente="PRESUPUESTO_2026")
        call_command("corregir_pronostico_nombres_point", stdout=StringIO())
        self.assertEqual(PronosticoVenta.objects.filter(receta=receta).count(), 0)


class ReclasificarInversionTests(TestCase):
    """Rubros de inversión se mueven a CAPEX; las refacciones se quedan."""

    def test_mueve_inversion_y_respeta_refacciones(self):
        gv = AreaPresupuesto.objects.create(nombre="Gastos venta", codigo="gastos-venta")
        AreaPresupuesto.objects.create(nombre="CAPEX", codigo="capex")
        inversion = RubroPresupuesto.objects.create(
            area=gv, concepto="Apertura sucursal", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        refaccion = RubroPresupuesto.objects.create(
            area=gv, concepto="Refrigerador 1", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        call_command("reclasificar_inversion_capex", stdout=StringIO())
        inversion.refresh_from_db()
        refaccion.refresh_from_db()
        self.assertEqual(inversion.area.codigo, "capex")
        self.assertEqual(inversion.metadata.get("area_anterior"), "gastos-venta")
        self.assertEqual(refaccion.area.codigo, "gastos-venta")

    def test_dry_run_no_mueve(self):
        gv = AreaPresupuesto.objects.create(nombre="Gastos venta", codigo="gastos-venta")
        AreaPresupuesto.objects.create(nombre="CAPEX", codigo="capex")
        r = RubroPresupuesto.objects.create(
            area=gv, concepto="Apertura sucursal", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        call_command("reclasificar_inversion_capex", "--dry-run", stdout=StringIO())
        r.refresh_from_db()
        self.assertEqual(r.area.codigo, "gastos-venta")


class ConsolidacionNocturnaVentanaTests(TestCase):
    """La nocturna barre el mes corriente + 2 anteriores (fuentes tardías)."""

    def test_sin_argumento_cubre_tres_meses(self):
        from unittest.mock import patch

        from reportes.tasks import task_consolidar_presupuesto_real

        consolidados = []

        class FakeService:
            def consolidar(self, periodo):
                consolidados.append(periodo)

                class R:
                    def as_dict(self):
                        return {"periodo": str(periodo)}

                return R()

        with patch(
            "reportes.services_presupuesto_real.PresupuestoRealConsolidacionService",
            FakeService,
        ):
            task_consolidar_presupuesto_real.run()
        self.assertEqual(len(consolidados), 3)
        self.assertEqual(len({p for p in consolidados}), 3)
        self.assertTrue(all(p.day == 1 for p in consolidados))


class DesgloseConceptosTests(TestCase):
    """El desglose por concepto suma sucursales y agrupa por bloque del P&L."""

    def test_concepto_suma_sucursales(self):
        from django.contrib.auth import get_user_model

        gv = AreaPresupuesto.objects.create(nombre="Gastos venta", codigo="gastos-venta")
        s1 = Sucursal.objects.create(nombre="Uno", codigo="DES-S1")
        s2 = Sucursal.objects.create(nombre="Dos", codigo="DES-S2")
        for suc, real in ((s1, "100"), (s2, "150")):
            rubro = RubroPresupuesto.objects.create(
                area=gv, concepto="Renta", sucursal=suc, tipo=RubroPresupuesto.TIPO_EGRESO
            )
            LineaPresupuestoMensual.objects.create(
                rubro=rubro, periodo=date(2026, 2, 1),
                monto_presupuesto=Decimal("120"), monto_real=Decimal(real),
                fuente_real="AUTO:GASTO_OPERATIVO",
            )
        user = get_user_model().objects.create_superuser("dg_desglose", "dgd@test.mx", "clave")
        self.client.force_login(user)
        response = self.client.get("/reportes/estado-resultados/?year=2026")
        desglose = response.context["desglose"]
        bloque_gv = next(b for b in desglose if b["grupo"] == "Gastos de venta")
        renta = next(f for f in bloque_gv["conceptos"] if f["label"] == "Renta")
        self.assertEqual(renta["meses"][1]["real"], Decimal("250"))
        self.assertEqual(renta["meses"][1]["ppto"], Decimal("240"))


class MantenimientoUnidadTests(TestCase):
    """La fuente MANTENIMIENTO_UNIDAD liga rubros de flotilla con ReporteUnidad."""

    def test_costo_servicio_fluye_al_rubro(self):
        from logistica.models import Repartidor, ReporteUnidad, Unidad

        area = AreaPresupuesto.objects.create(nombre="Logística", codigo="logistica")
        rubro = RubroPresupuesto.objects.create(
            area=area, concepto="Peugeot Partner", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        linea = LineaPresupuestoMensual.objects.create(
            rubro=rubro, periodo=date(2026, 7, 1), monto_presupuesto=Decimal("5000"),
        )
        suc = Sucursal.objects.create(nombre="Matriz MU", codigo="MU-MTZ")
        unidad = Unidad.objects.create(codigo="GS-P1", descripcion="Peugeot Partner", sucursal=suc)
        ReporteUnidad.objects.create(
            unidad=unidad, tipo="FALLA", severidad="MEDIA", descripcion="Servicio frenos",
            costo_servicio=Decimal("12287"), fecha_reporte=timezone.now().replace(year=2026, month=7, day=10),
        )
        ReglaFuenteRubro.objects.create(
            rubro=rubro, tipo_fuente=ReglaFuenteRubro.FUENTE_MANTENIMIENTO_UNIDAD,
            filtros={"unidad_codigo": "GS-P1"},
        )
        PresupuestoRealConsolidacionService().consolidar(periodo=date(2026, 7, 1))
        linea.refresh_from_db()
        self.assertEqual(linea.monto_real, Decimal("12287.00"))
        self.assertEqual(linea.fuente_real, "AUTO:MANTENIMIENTO_UNIDAD")


class CombustibleYMantEquipoTests(TestCase):
    """Combustible por unidad y mantenimiento de equipos fluyen desde julio."""

    def test_combustible_suma_unidades_y_respeta_desde(self):
        from logistica.models import CargaCombustibleUnidad, Repartidor, Unidad

        area = AreaPresupuesto.objects.create(nombre="Logística", codigo="logistica")
        rubro = RubroPresupuesto.objects.create(
            area=area, concepto="Diesel", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        for mes, real_previo in ((6, "18000"), (7, None)):
            LineaPresupuestoMensual.objects.create(
                rubro=rubro, periodo=date(2026, mes, 1), monto_presupuesto=Decimal("22000"),
                monto_real=Decimal(real_previo) if real_previo else None,
                fuente_real="AUTO:LEGADO" if real_previo else "",
            )
        suc = Sucursal.objects.create(nombre="Matriz CB", codigo="CB-MTZ")
        u1 = Unidad.objects.create(codigo="GS-DC1", descripcion="Fiat Ducato", sucursal=suc)
        u2 = Unidad.objects.create(codigo="GS-PM1", descripcion="Peugeot Manager", sucursal=suc)
        from django.contrib.auth import get_user_model

        from logistica.models import BitacoraSalidaLlegada
        user_rep = get_user_model().objects.create_user("rep_cb", "rep@test.mx", "clave")
        rep = Repartidor.objects.create(user=user_rep, sucursal=suc)
        bit = BitacoraSalidaLlegada.objects.create(
            repartidor=rep, unidad=u1, km_salida=1, nivel_gas_salida="1/2",
        )
        for u, monto, mes in ((u1, "6600", 7), (u2, "7600", 7), (u1, "12800", 6)):
            carga = CargaCombustibleUnidad.objects.create(
                bitacora=bit, unidad=u, repartidor=rep, litros=Decimal("40"),
                importe_total=Decimal(monto),
            )
            # fecha_registro es auto_now_add: fijar el mes vía update
            CargaCombustibleUnidad.objects.filter(pk=carga.pk).update(
                fecha_registro=timezone.now().replace(year=2026, month=mes, day=5)
            )
        ReglaFuenteRubro.objects.create(
            rubro=rubro, tipo_fuente=ReglaFuenteRubro.FUENTE_COMBUSTIBLE_UNIDAD,
            filtros={"unidades": ["GS-DC1", "GS-PM1"], "desde": "2026-07"},
        )
        PresupuestoRealConsolidacionService().consolidar(periodo=date(2026, 7, 1))
        PresupuestoRealConsolidacionService().consolidar(periodo=date(2026, 6, 1))
        jul = LineaPresupuestoMensual.objects.get(rubro=rubro, periodo=date(2026, 7, 1))
        jun = LineaPresupuestoMensual.objects.get(rubro=rubro, periodo=date(2026, 6, 1))
        self.assertEqual(jul.monto_real, Decimal("14200.00"))
        self.assertEqual(jul.fuente_real, "AUTO:COMBUSTIBLE_UNIDAD")
        # junio: antes del 'desde' → conserva el legado del Excel
        self.assertEqual(jun.monto_real, Decimal("18000"))
        self.assertEqual(jun.fuente_real, "AUTO:LEGADO")

    def test_mantenimiento_equipo_produccion(self):
        from activos.models import Activo, OrdenMantenimiento

        area = AreaPresupuesto.objects.create(nombre="Producción", codigo="produccion")
        rubro = RubroPresupuesto.objects.create(
            area=area, concepto="Mantenimiento equipo/maquinaria", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        LineaPresupuestoMensual.objects.create(
            rubro=rubro, periodo=date(2026, 7, 1), monto_presupuesto=Decimal("7000"),
        )
        horno = Activo.objects.create(nombre="Horno Test", ubicacion="HORNOS")
        vitrina = Activo.objects.create(nombre="Vitrina Test", ubicacion="LEYVA")
        for activo, costo in ((horno, "1800"), (vitrina, "999")):
            OrdenMantenimiento.objects.create(
                activo_ref=activo, tipo="CORRECTIVO", prioridad="MEDIA", estatus="CERRADA",
                descripcion="test", costo_repuestos=Decimal(costo), costo_mano_obra=0, costo_otros=0,
            )
        ReglaFuenteRubro.objects.create(
            rubro=rubro, tipo_fuente=ReglaFuenteRubro.FUENTE_MANTENIMIENTO_EQUIPO,
            filtros={"ubicaciones_produccion": True, "desde": "2026-07"},
        )
        PresupuestoRealConsolidacionService().consolidar(periodo=date(2026, 7, 1))
        linea = LineaPresupuestoMensual.objects.get(rubro=rubro, periodo=date(2026, 7, 1))
        # solo el horno (HORNOS es producción); la vitrina de LEYVA no
        self.assertEqual(linea.monto_real, Decimal("1800.00"))
        self.assertEqual(linea.fuente_real, "AUTO:MANTENIMIENTO_EQUIPO")


class MantEquipoPorSucursalTests(TestCase):
    """La regla por_sucursal hereda la sucursal del rubro."""

    def test_cada_sucursal_toma_sus_ordenes(self):
        from activos.models import Activo, OrdenMantenimiento

        area = AreaPresupuesto.objects.create(nombre="Gastos venta", codigo="gastos-venta")
        s1 = Sucursal.objects.create(nombre="Matriz ME", codigo="ME-MTZ")
        s2 = Sucursal.objects.create(nombre="Leyva ME", codigo="ME-LEY")
        lineas = {}
        for suc in (s1, s2):
            rubro = RubroPresupuesto.objects.create(
                area=area, concepto="Mantenimiento equipo/maquinaria",
                sucursal=suc, tipo=RubroPresupuesto.TIPO_EGRESO,
            )
            lineas[suc.codigo] = LineaPresupuestoMensual.objects.create(
                rubro=rubro, periodo=date(2026, 7, 1), monto_presupuesto=Decimal("1000"),
            )
            ReglaFuenteRubro.objects.create(
                rubro=rubro, tipo_fuente=ReglaFuenteRubro.FUENTE_MANTENIMIENTO_EQUIPO,
                filtros={"por_sucursal": True, "desde": "2026-07"},
            )
        a1 = Activo.objects.create(nombre="Refri ME1", ubicacion="MATRIZ", sucursal=s1)
        a2 = Activo.objects.create(nombre="Vitrina ME2", ubicacion="LEYVA", sucursal=s2)
        for activo, costo in ((a1, "900"), (a2, "400")):
            OrdenMantenimiento.objects.create(
                activo_ref=activo, tipo="CORRECTIVO", prioridad="MEDIA", estatus="CERRADA",
                descripcion="t", costo_repuestos=Decimal(costo), costo_mano_obra=0, costo_otros=0,
            )
        PresupuestoRealConsolidacionService().consolidar(periodo=date(2026, 7, 1))
        for codigo, esperado in (("ME-MTZ", "900.00"), ("ME-LEY", "400.00")):
            linea = LineaPresupuestoMensual.objects.get(pk=lineas[codigo].pk)
            self.assertEqual(linea.monto_real, Decimal(esperado))
            self.assertEqual(linea.fuente_real, "AUTO:MANTENIMIENTO_EQUIPO")


class ComplementosClasificadosTests(TestCase):
    """Complementos = productos del catálogo curado; postres los excluye."""

    def test_separacion_sin_doble_conteo(self):
        from pos_bridge.models import PointBranch, PointProductCategory

        area = AreaPresupuesto.objects.create(nombre="Resultados", codigo="resultados")
        postres = RubroPresupuesto.objects.create(
            area=area, concepto="Venta postres", tipo=RubroPresupuesto.TIPO_INGRESO
        )
        compl = RubroPresupuesto.objects.create(
            area=area, concepto="Venta complementos", tipo=RubroPresupuesto.TIPO_INGRESO
        )
        for rubro in (postres, compl):
            LineaPresupuestoMensual.objects.create(
                rubro=rubro, periodo=date(2026, 6, 1), monto_presupuesto=Decimal("100"),
            )
        PointProductCategory.objects.create(codigo_point="V1", nombre="Vela Granmark", category="SERVICIO_ACCESORIO")
        branch = PointBranch.objects.create(external_id="cc-br", name="Matriz CC")
        for producto, monto in (("Pastel X", "1000"), ("Vela Granmark", "50")):
            PointSalesDailyProductFact.objects.create(
                branch=branch, sale_date=date(2026, 6, 5), sucursal_nombre="Matriz CC",
                categoria="Cat", producto_nombre_historico=producto,
                total_venta=Decimal(monto), total_venta_neta=Decimal(monto),
            )
        ReglaFuenteRubro.objects.create(
            rubro=postres, tipo_fuente=ReglaFuenteRubro.FUENTE_VENTA_POS,
            filtros={"total_empresa": True, "excluir_clasificados": True},
        )
        ReglaFuenteRubro.objects.create(
            rubro=compl, tipo_fuente=ReglaFuenteRubro.FUENTE_VENTA_POS,
            filtros={"clasificacion_catalogo": ["REVENTA", "SERVICIO_ACCESORIO", "TOPPING"]},
        )
        PresupuestoRealConsolidacionService().consolidar(periodo=date(2026, 6, 1))
        lp = LineaPresupuestoMensual.objects.get(rubro=postres)
        lc = LineaPresupuestoMensual.objects.get(rubro=compl)
        self.assertEqual(lp.monto_real, Decimal("1000.00"))
        self.assertEqual(lc.monto_real, Decimal("50.00"))
        # suma = total Point; nada doble, nada perdido
        self.assertEqual(lp.monto_real + lc.monto_real, Decimal("1050.00"))


class CostoReventaComplementosTests(TestCase):
    """Costos complementos = unidades vendidas × costo de reventa, desde junio."""

    def test_unidades_por_costo_con_desde(self):
        from pos_bridge.models import PointBranch, PointProduct, PointProductCategory
        from reportes.models import ProductoReventaCosto

        area = AreaPresupuesto.objects.create(nombre="Resultados", codigo="resultados")
        rubro = RubroPresupuesto.objects.create(
            area=area, concepto="Costos complementos", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        for mes, legado in ((5, "9000"), (6, None)):
            LineaPresupuestoMensual.objects.create(
                rubro=rubro, periodo=date(2026, mes, 1), monto_presupuesto=Decimal("10000"),
                monto_real=Decimal(legado) if legado else None,
                fuente_real="AUTO:LEGADO" if legado else "",
            )
        PointProductCategory.objects.create(codigo_point="CR1", nombre="Vela Magica", category="SERVICIO_ACCESORIO")
        producto = PointProduct.objects.create(external_id="cr-1", sku="CR1", name="Vela Magica")
        ProductoReventaCosto.objects.create(
            producto_point=producto, costo_unitario=Decimal("4.50"),
            fecha_vigencia=date(2026, 1, 10), fuente="TEST", source_hash="cr-hash-1",
        )
        branch = PointBranch.objects.create(external_id="cr-br", name="Matriz CR")
        for mes, unidades in ((5, "100"), (6, "200")):
            PointSalesDailyProductFact.objects.create(
                branch=branch, sale_date=date(2026, mes, 10), sucursal_nombre="Matriz CR",
                categoria="Accesorios", producto_nombre_historico="Vela Magica",
                point_product=producto, total_cantidad=Decimal(unidades),
                total_venta=Decimal("1"), total_venta_neta=Decimal("1"),
            )
        ReglaFuenteRubro.objects.create(
            rubro=rubro, tipo_fuente=ReglaFuenteRubro.FUENTE_COSTO_REVENTA,
            filtros={"desde": "2026-06"},
        )
        PresupuestoRealConsolidacionService().consolidar(periodo=date(2026, 6, 1))
        PresupuestoRealConsolidacionService().consolidar(periodo=date(2026, 5, 1))
        jun = LineaPresupuestoMensual.objects.get(rubro=rubro, periodo=date(2026, 6, 1))
        may = LineaPresupuestoMensual.objects.get(rubro=rubro, periodo=date(2026, 5, 1))
        self.assertEqual(jun.monto_real, Decimal("900.00"))  # 200 × 4.50
        self.assertEqual(jun.fuente_real, "AUTO:COSTO_REVENTA")
        self.assertEqual(may.monto_real, Decimal("9000"))  # legado intacto
        self.assertEqual(may.fuente_real, "AUTO:LEGADO")


class MermaProductoTests(TestCase):
    """La fuente MERMA_PRODUCTO valúa la merma Point (MermaPOS) a costo de receta."""

    def test_merma_point_valuada_respeta_desde(self):
        from control.models import MermaPOS
        from recetas.models import LineaReceta, Receta

        area = AreaPresupuesto.objects.create(nombre="Resultados (P&L)", codigo="resultados")
        rubro = RubroPresupuesto.objects.create(
            area=area, concepto="Merma", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        jul = LineaPresupuestoMensual.objects.create(
            rubro=rubro, periodo=date(2026, 7, 1), monto_presupuesto=Decimal("2000"),
        )
        may = LineaPresupuestoMensual.objects.create(
            rubro=rubro, periodo=date(2026, 5, 1), monto_presupuesto=Decimal("2000"),
            monto_real=Decimal("1495.59"), fuente_real="AUTO:LEGADO",
        )
        # Receta con costo total 100 y rendimiento 10 → costo unitario 10.
        receta = Receta.objects.create(
            nombre="Pastel Merma Test", hash_contenido="t-merma", rendimiento_cantidad=10
        )
        LineaReceta.objects.create(
            receta=receta, insumo_texto="base", costo_linea_excel=Decimal("100"),
            match_status=LineaReceta.STATUS_AUTO,
        )
        suc = Sucursal.objects.create(nombre="Matriz Merma", codigo="MM-MTZ")
        # Sucursal y CEDIS suman igual: 3 × 10 + 2 × 10.
        MermaPOS.objects.create(receta=receta, sucursal=suc, fecha=date(2026, 7, 5), cantidad=Decimal("3"))
        MermaPOS.objects.create(receta=receta, fecha=date(2026, 7, 6), cantidad=Decimal("2"))
        # Sin receta ligada: no se puede valuar, no debe tronar.
        MermaPOS.objects.create(producto_texto="gelatina suelta", fecha=date(2026, 7, 6), cantidad=Decimal("1"))
        # Producto terminado sin rendimiento: el costo total es por pieza.
        producto = Receta.objects.create(nombre="Piñatero Merma Test", hash_contenido="t-merma-p")
        LineaReceta.objects.create(
            receta=producto, insumo_texto="base", costo_linea_excel=Decimal("40"),
            match_status=LineaReceta.STATUS_AUTO,
        )
        MermaPOS.objects.create(receta=producto, fecha=date(2026, 7, 8), cantidad=Decimal("1"))
        ReglaFuenteRubro.objects.create(
            rubro=rubro, tipo_fuente=ReglaFuenteRubro.FUENTE_MERMA_PRODUCTO,
            filtros={"desde": "2026-06"},
        )
        service = PresupuestoRealConsolidacionService()
        service.consolidar(periodo=date(2026, 7, 1))
        service.consolidar(periodo=date(2026, 5, 1))
        jul.refresh_from_db()
        may.refresh_from_db()
        self.assertEqual(jul.monto_real, Decimal("90.00"))  # (3 + 2) × 10 + 1 × 40
        self.assertEqual(jul.fuente_real, "AUTO:MERMA_PRODUCTO")
        self.assertEqual(may.monto_real, Decimal("1495.59"))  # legado intacto
        self.assertEqual(may.fuente_real, "AUTO:LEGADO")


class ImportNominaHojaGeneralTests(TestCase):
    """El archivo de nómina alimenta el área de control SOLO con la hoja GENERAL."""

    def test_nomina_toma_general_e_ignora_departamentos(self):
        import tempfile

        from openpyxl import Workbook

        from reportes.services_presupuesto_maestro import PresupuestoMaestroImportService

        wb = Workbook()
        general = wb.active
        general.title = "GENERAL"
        general.append(["CUENTA", "DESCRIPCION", "ENERO", "", ""])
        general.append(["", "", "PRESUPUESTADO", "REAL", "VARIACION"])
        general.append(["", "SUELDO", 632149, "", ""])
        admin = wb.create_sheet("ADMINISTRACION")
        admin.append(["CUENTA", "DESCRIPCION", "ENERO", "", ""])
        admin.append(["", "", "PRESUPUESTADO", "REAL", "VARIACION"])
        admin.append(["", "SUELDO", 134197, "", ""])
        admin.append(["", "PLAYERA", 300, "", ""])
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)

        PresupuestoMaestroImportService().import_file(
            archivo=tmp.name, area_code="nomina", version="ORIGINAL", year=2026
        )
        linea = LineaPresupuestoMensual.objects.get(
            periodo=date(2026, 1, 1), rubro__concepto__iexact="sueldo",
            rubro__area__codigo="nomina",
        )
        # Gana la hoja GENERAL, no la última hoja de departamento.
        self.assertEqual(linea.monto_presupuesto, Decimal("632149"))
        # Los conceptos exclusivos de hojas de departamento no entran al control.
        self.assertFalse(
            RubroPresupuesto.objects.filter(
                area__codigo="nomina", concepto__iexact="playera"
            ).exists()
        )


class RecihoCompartidoPorcentajeTests(TestCase):
    """El recibo compartido Matriz+CEDIS se reparte por porcentaje de la regla."""

    def test_porcentaje_sobre_centro_compartido(self):
        from reportes.models import CentroCosto

        prod = AreaPresupuesto.objects.create(nombre="Producción", codigo="produccion")
        gv = AreaPresupuesto.objects.create(nombre="Gastos de Venta", codigo="gastos-venta")
        matriz = Sucursal.objects.create(nombre="Sucursal Matriz", codigo="RC-MTZ")
        r_prod = RubroPresupuesto.objects.create(
            area=prod, concepto="Energía eléctrica", tipo=RubroPresupuesto.TIPO_EGRESO
        )
        r_mtz = RubroPresupuesto.objects.create(
            area=gv, concepto="Energía eléctrica", sucursal=matriz,
            tipo=RubroPresupuesto.TIPO_EGRESO,
        )
        l_prod = LineaPresupuestoMensual.objects.create(
            rubro=r_prod, periodo=date(2026, 7, 1), monto_presupuesto=Decimal("30000"),
        )
        l_mtz = LineaPresupuestoMensual.objects.create(
            rubro=r_mtz, periodo=date(2026, 7, 1), monto_presupuesto=Decimal("15000"),
        )
        cat = CategoriaGasto.objects.create(
            codigo="LUZ_SUC", nombre="Luz y energía eléctrica",
            capa_objetivo=CategoriaGasto.CAPA_SUCURSAL,
        )
        centro = CentroCosto.objects.create(
            codigo="COMPARTIDO_MC", nombre="Compartido Matriz-CEDIS", tipo="COMPARTIDO"
        )
        GastoOperativoMensual.objects.create(
            periodo=date(2026, 7, 1), categoria_gasto=cat, centro_costo=centro,
            monto=Decimal("40000"), tipo_dato=GastoOperativoMensual.TIPO_DATO_REAL,
        )
        ReglaFuenteRubro.objects.create(
            rubro=r_prod, tipo_fuente=ReglaFuenteRubro.FUENTE_GASTO_OPERATIVO,
            categoria_gasto=cat, filtros={"centro_tipo": "COMPARTIDO", "porcentaje": 65},
        )
        ReglaFuenteRubro.objects.create(
            rubro=r_mtz, tipo_fuente=ReglaFuenteRubro.FUENTE_GASTO_OPERATIVO,
            categoria_gasto=cat, filtros={"centro_tipo": "COMPARTIDO", "porcentaje": 35},
        )
        PresupuestoRealConsolidacionService().consolidar(periodo=date(2026, 7, 1))
        l_prod.refresh_from_db()
        l_mtz.refresh_from_db()
        self.assertEqual(l_prod.monto_real, Decimal("26000.00"))  # 65%
        self.assertEqual(l_mtz.monto_real, Decimal("14000.00"))  # 35%


class EmpaquesGastoOperativoTests(TestCase):
    """Los rubros "Etiquetas, bolsas, cajas y empaques" se llenan de
    GASTO_OPERATIVO/EMPAQUE: por sucursal con herencia del rubro y el de
    administración solo con centros CORPORATIVO. NUNCA de CONSUMO_MP: los
    insumos EMPAQUE ya generan filas en ConsumoInsumoMensual y el renglón
    Costos del P&L (total_empresa) las suma — duplicaría a nivel empresa."""

    CONCEPTO = "Etiquetas, bolsas, cajas y empaques"

    @classmethod
    def setUpTestData(cls):
        CategoriaGasto.objects.create(
            codigo="RENTA", nombre="Renta sucursal", capa_objetivo=CategoriaGasto.CAPA_EMPRESA
        )
        cls.gv = AreaPresupuesto.objects.create(nombre="Gastos de Venta", codigo="gastos-venta")
        cls.admin = AreaPresupuesto.objects.create(nombre="Administración", codigo="administracion")
        cls.sucursal = Sucursal.objects.create(codigo="EMP01", nombre="Sucursal Centro")
        cls.r_suc = RubroPresupuesto.objects.create(
            area=cls.gv, concepto=cls.CONCEPTO, sucursal=cls.sucursal,
            tipo=RubroPresupuesto.TIPO_EGRESO,
        )
        cls.r_admin = RubroPresupuesto.objects.create(
            area=cls.admin, concepto=cls.CONCEPTO, tipo=RubroPresupuesto.TIPO_EGRESO
        )
        cls.periodo = date(2026, 6, 1)
        cls.l_suc = LineaPresupuestoMensual.objects.create(
            rubro=cls.r_suc, periodo=cls.periodo, monto_presupuesto=Decimal("2000"),
        )
        cls.l_admin = LineaPresupuestoMensual.objects.create(
            rubro=cls.r_admin, periodo=cls.periodo, monto_presupuesto=Decimal("20000"),
            monto_real=Decimal("19000"), fuente_real="AUTO:LEGADO",
        )
        call_command("seed_reglas_fuente_rubro", stdout=StringIO())

    def test_seed_crea_reglas_y_autoprovisiona_categoria_empaque(self):
        cat = CategoriaGasto.objects.get(codigo="EMPAQUE")
        regla_suc = ReglaFuenteRubro.objects.get(rubro=self.r_suc)
        self.assertEqual(regla_suc.tipo_fuente, ReglaFuenteRubro.FUENTE_GASTO_OPERATIVO)
        self.assertEqual(regla_suc.categoria_gasto, cat)
        self.assertNotIn("centro_tipo", regla_suc.filtros or {})
        regla_admin = ReglaFuenteRubro.objects.get(rubro=self.r_admin)
        self.assertEqual(regla_admin.categoria_gasto, cat)
        self.assertEqual((regla_admin.filtros or {}).get("centro_tipo"), "CORPORATIVO")

    def test_sucursal_hereda_y_admin_solo_toma_corporativo(self):
        cat = CategoriaGasto.objects.get(codigo="EMPAQUE")
        centro_suc = CentroCosto.objects.create(
            codigo="EMP_SUC", nombre="Centro", tipo="SUCURSAL_VENTA", sucursal=self.sucursal
        )
        centro_corp = CentroCosto.objects.create(
            codigo="EMP_CORP", nombre="Oficinas", tipo="CORPORATIVO"
        )
        GastoOperativoMensual.objects.create(
            periodo=self.periodo, categoria_gasto=cat, centro_costo=centro_suc,
            monto=Decimal("1800"), tipo_dato=GastoOperativoMensual.TIPO_DATO_REAL,
        )
        GastoOperativoMensual.objects.create(
            periodo=self.periodo, categoria_gasto=cat, centro_costo=centro_corp,
            monto=Decimal("500"), tipo_dato=GastoOperativoMensual.TIPO_DATO_REAL,
        )
        PresupuestoRealConsolidacionService().consolidar(periodo=self.periodo)
        self.l_suc.refresh_from_db()
        self.l_admin.refresh_from_db()
        self.assertEqual(self.l_suc.monto_real, Decimal("1800.00"))
        self.assertEqual(self.l_suc.fuente_real, "AUTO:GASTO_OPERATIVO")
        self.assertEqual(self.l_admin.monto_real, Decimal("500.00"))

    def test_admin_sin_captura_corporativa_conserva_legado(self):
        PresupuestoRealConsolidacionService().consolidar(periodo=self.periodo)
        self.l_admin.refresh_from_db()
        self.assertEqual(self.l_admin.monto_real, Decimal("19000.00"))
        self.assertEqual(self.l_admin.fuente_real, "AUTO:LEGADO")
        self.assertTrue((self.l_admin.metadata or {}).get("sin_datos_fuente"))


class ReestructuraBebidasOtrosTemporadaTests(TestCase):
    """Separación del renglón BEBIDAS/OTROS · ESPECIAL/TEMPORADA (2026-07-18)."""

    @classmethod
    def setUpTestData(cls):
        cls.area = AreaPresupuesto.objects.create(nombre="Ventas", codigo="ventas")

        def rubro(concepto, lineas):
            r = RubroPresupuesto.objects.create(
                area=cls.area, concepto=concepto, tipo=RubroPresupuesto.TIPO_INGRESO
            )
            for periodo, ppto, real, fuente in lineas:
                LineaPresupuestoMensual.objects.create(
                    rubro=r,
                    periodo=periodo,
                    monto_presupuesto=Decimal(ppto),
                    monto_real=Decimal(real) if real is not None else None,
                    fuente_real=fuente,
                )
            return r

        jun, jul = date(2026, 6, 1), date(2026, 7, 1)
        cls.mixto = rubro(
            "BEBIDAS/OTROS · ESPECIAL/TEMPORADA",
            [(jun, "152475", "448", "AUTO:LEGADO"), (jul, "424400", None, "")],
        )
        cls.galleta = rubro(
            "GALLETA · ESPECIAL/TEMPORADA",
            [(jun, "1015", "2030", "AUTO:LEGADO"), (jul, "100", "50", "MANUAL:paula")],
        )
        cls.coca = rubro("Coca-cola", [(jun, "3200", "1300", "AUTO:VENTA_POS")])
        cls.te = rubro("TE", [(jun, "2800", "950", "AUTO:VENTA_POS")])
        ReglaFuenteRubro.objects.create(
            rubro=cls.coca,
            tipo_fuente=ReglaFuenteRubro.FUENTE_VENTA_POS,
            filtros={"categoria_pos": "Coca-cola"},
        )

    def test_reestructura_fusiona_renombra_y_crea_otros(self):
        salida = StringIO()
        call_command("reestructurar_bebidas_otros_temporada", stdout=salida)

        temporada = RubroPresupuesto.objects.get(pk=self.mixto.pk)
        self.assertEqual(temporada.concepto, "Especiales/Temporada")
        self.assertEqual(
            temporada.metadata["nombre_excel"], "BEBIDAS/OTROS · ESPECIAL/TEMPORADA"
        )
        jun = temporada.lineas_mensuales.get(periodo=date(2026, 6, 1))
        # Presupuestos sumados y reales AUTO/legado sumados.
        self.assertEqual(jun.monto_presupuesto, Decimal("153490"))
        self.assertEqual(jun.monto_real, Decimal("2478"))
        # El real MANUAL del origen NO se fusiona (nunca se pisa ni se mueve).
        jul = temporada.lineas_mensuales.get(periodo=date(2026, 7, 1))
        self.assertEqual(jul.monto_presupuesto, Decimal("424500"))
        self.assertIsNone(jul.monto_real)
        self.assertIn("MANUAL", salida.getvalue())

        bebidas = RubroPresupuesto.objects.get(pk=self.coca.pk)
        self.assertEqual(bebidas.concepto, "Bebidas")
        jun_beb = bebidas.lineas_mensuales.get(periodo=date(2026, 6, 1))
        self.assertEqual(jun_beb.monto_presupuesto, Decimal("6000"))
        self.assertEqual(jun_beb.monto_real, Decimal("2250"))

        for origen in (self.galleta, self.te):
            origen.refresh_from_db()
            self.assertFalse(origen.activo)
            self.assertIn("fusionado", origen.metadata["motivo_desactivacion"])

        otros = RubroPresupuesto.objects.get(area=self.area, concepto="Otros")
        self.assertEqual(otros.lineas_mensuales.count(), 2)
        for linea in otros.lineas_mensuales.all():
            self.assertEqual(linea.monto_presupuesto, Decimal("0"))

    def test_es_idempotente(self):
        call_command("reestructurar_bebidas_otros_temporada", stdout=StringIO())
        call_command("reestructurar_bebidas_otros_temporada", stdout=StringIO())
        temporada = RubroPresupuesto.objects.get(pk=self.mixto.pk)
        jun = temporada.lineas_mensuales.get(periodo=date(2026, 6, 1))
        self.assertEqual(jun.monto_presupuesto, Decimal("153490"))
        self.assertEqual(
            RubroPresupuesto.objects.filter(area=self.area, concepto="Otros").count(), 1
        )


class ExclusividadOverridesSeedTests(TestCase):
    """Los productos/categorías reclamados por overrides del CSV no pueden
    ganarse por matching difuso (ningún producto Point en dos reglas)."""

    @classmethod
    def setUpTestData(cls):
        CategoriaGasto.objects.create(
            codigo="RENTA", nombre="Renta sucursal", capa_objetivo=CategoriaGasto.CAPA_EMPRESA
        )
        cls.area = AreaPresupuesto.objects.create(nombre="Ventas", codigo="ventas")
        branch = PointBranch.objects.create(external_id="EXC-BR", name="Centro")
        for cat, prod in [
            ("Pastel Mediano", "Pastel Lotus Mediano"),
            ("Café", "Capuchino"),
            ("Coca-cola", "COCA-COLA 450 ML"),
            ("TE", "TE DEL JARDIN"),
            ("Otros postres", "Cubeta de Crema Para Fresas"),
        ]:
            PointSalesDailyProductFact.objects.create(
                branch=branch, sale_date=date(2026, 6, 3), sucursal_nombre="Centro",
                categoria=cat, producto_nombre_historico=prod,
                total_venta=Decimal("10"), total_venta_neta=Decimal("9"),
            )

        def rubro(concepto):
            r = RubroPresupuesto.objects.create(
                area=cls.area, concepto=concepto, tipo=RubroPresupuesto.TIPO_INGRESO
            )
            LineaPresupuestoMensual.objects.create(
                rubro=r, periodo=date(2026, 6, 1), monto_presupuesto=Decimal("1")
            )
            return r

        cls.temporada = rubro("Especiales/Temporada")
        cls.bebidas = rubro("Bebidas")
        cls.otros = rubro("Otros")
        cls.competidor = rubro("PASTEL MEDIANO · LOTUS")

    def test_override_saca_sus_productos_y_categorias_del_pool_difuso(self):
        call_command("seed_reglas_fuente_rubro", stdout=StringIO())

        temporada = ReglaFuenteRubro.objects.get(rubro=self.temporada)
        self.assertIn("Pastel Lotus Mediano", temporada.filtros["productos_pos"])

        # El rubro que el matching difuso habría cruzado con "Pastel Lotus
        # Mediano" (score 100) queda SIN asignación: el override manda.
        competidor = ReglaFuenteRubro.objects.get(rubro=self.competidor)
        self.assertNotIn("productos_pos", competidor.filtros)
        self.assertNotIn("categoria_pos", competidor.filtros)

        otros = ReglaFuenteRubro.objects.get(rubro=self.otros)
        self.assertEqual(otros.filtros.get("categoria_pos"), "Otros postres")
        self.assertEqual(ReglaFuenteRubro.objects.filter(rubro=self.bebidas).count(), 3)

    def test_consolidacion_suma_las_tres_categorias_de_bebidas(self):
        call_command("seed_reglas_fuente_rubro", stdout=StringIO())
        PresupuestoRealConsolidacionService().consolidar(periodo=date(2026, 6, 1))
        linea = self.bebidas.lineas_mensuales.get(periodo=date(2026, 6, 1))
        # Coca-cola (10) + TE (10) + Café (10); el resto no le pertenece.
        self.assertEqual(linea.monto_real, Decimal("30.00"))
        self.assertEqual(linea.fuente_real, "AUTO:VENTA_POS")


class CrearRenglonesFamiliasVentasTests(TestCase):
    """Renglones ppto-0 para familias vivas nunca presupuestadas (2026-07-19)."""

    def test_crea_renglones_idempotente(self):
        area = AreaPresupuesto.objects.create(nombre="Ventas", codigo="ventas")
        molde = RubroPresupuesto.objects.create(
            area=area, concepto="Especiales/Temporada", tipo=RubroPresupuesto.TIPO_INGRESO
        )
        for mes in (6, 7):
            LineaPresupuestoMensual.objects.create(rubro=molde, periodo=date(2026, mes, 1))
        call_command("crear_renglones_familias_ventas", stdout=StringIO())
        call_command("crear_renglones_familias_ventas", stdout=StringIO())

        granmark = RubroPresupuesto.objects.get(area=area, concepto="Granmark", activo=True)
        self.assertEqual(granmark.lineas_mensuales.count(), 2)
        for linea in granmark.lineas_mensuales.all():
            self.assertEqual(linea.monto_presupuesto, Decimal("0"))
        self.assertEqual(
            RubroPresupuesto.objects.filter(area=area, concepto="Piñatero Mini").count(), 1
        )


class ConciliacionCombustibleTests(TestCase):
    """El render de conciliación marca montos no redondos sin clasificar."""

    def test_monto_no_redondo_se_marca_revisar(self):
        from decimal import Decimal as D

        from reportes.services_conciliacion_combustible import _es_monto_redondo

        self.assertTrue(_es_monto_redondo(D("3000")))
        self.assertTrue(_es_monto_redondo(D("17600.00")))
        self.assertFalse(_es_monto_redondo(D("1691.02")))
        self.assertFalse(_es_monto_redondo(D("1250")))  # no múltiplo de 100
