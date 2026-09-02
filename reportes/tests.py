import csv
import re
from django.contrib.auth.models import Group, User
from django.core.cache import cache
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO
from unittest.mock import patch

from openpyxl import load_workbook

from maestros.models import CostoInsumo
from compras.models import OrdenCompra
from crm.models import Cliente, PedidoCliente
from inventario.models import ExistenciaInsumo, MovimientoInventario
from maestros.models import Insumo, Proveedor, UnidadMedida
from pos_bridge.models import (
    PointBranch,
    PointDailyBranchIndicator,
    PointDailySale,
    PointExtractionLog,
    PointInsumoInventorySnapshot,
    PointInventorySnapshot,
    PointSalesDailyCategoryFact,
    PointSalesDailyProductFact,
    PointMonthlySalesOfficial,
    PointProduct,
    PointSyncJob,
)
from recetas.models import (
    LineaReceta,
    PlanProduccion,
    PlanProduccionItem,
    ProductoMonthClosure,
    ProductoMonthClosureLine,
    Receta,
    RecetaAgrupacionAddon,
    RecetaCostoSemanal,
    VentaHistorica,
)
from core.views import _point_sales_month_total as core_point_sales_month_total
from reportes.executive_panels import (
    build_branch_contribution_panel,
    build_central_flow_panel,
    build_monthly_yoy_panel,
    build_sales_forecast_panel,
)
from reportes.dashboard_sales_dataset import get_dashboard_sales_dataset
from reportes.bi_utils import compute_bi_snapshot
from reportes.views import _point_sales_month_total as reportes_point_sales_month_total
from reportes.views import (
    _bi_daily_sales_snapshot,
    _bi_branch_weekday_comparisons,
    _bi_product_weekday_comparisons,
    _bi_production_summary,
    _faltantes_rows,
    _bi_inventory_snapshot,
    _sales_previous_dates,
    _sales_source_context,
    _ventas_historicas_bi_summary,
    _product_closure_month_options,
)
from pos_bridge.services.sales_matching_service import PointSalesMatchingService
from reportes.models import (
    CorteOficialDiario,
    EmpresaResultadoMensual,
    FactVentaDiaria,
    PresupuestoImport,
    PresupuestoLineaMensual,
    PresupuestoResumenMensual,
    ProductoPricingDecisionMensual,
    ProductoSucursalContribucionMensual,
)
from ventas.models import VentaAutoritativaPoint
from reportes import executive_panels as executive_panels_module
from ventas.services import sales_truth as sales_truth_module


def _clear_reportes_runtime_caches() -> None:
    """Aísla tests que comparten proceso.

    Los bumps de versión del cache Django corren con transaction.on_commit y
    nunca se disparan dentro de TestCase, y los cortes de venta usan
    lru_cache a nivel de módulo (executive_panels, sales_truth); ambos deben
    limpiarse para que un test no lea datos del test anterior.
    """
    cache.clear()
    for module in (executive_panels_module, sales_truth_module):
        for attr in vars(module).values():
            if callable(attr) and hasattr(attr, "cache_clear"):
                attr.cache_clear()


class ReportesBITests(TestCase):
    def setUp(self):
        _clear_reportes_runtime_caches()
        self.user = User.objects.create_user(username="lectura_reportes", password="pass123")
        group, _ = Group.objects.get_or_create(name="LECTURA")
        self.user.groups.add(group)
        self.client.login(username="lectura_reportes", password="pass123")

        cliente = Cliente.objects.create(nombre="Cliente BI")
        PedidoCliente.objects.create(cliente=cliente, descripcion="Pedido BI", monto_estimado=1200)
        prov = Proveedor.objects.create(nombre="Proveedor BI")
        solicitud_insumo = None
        # Orden sin solicitud para no depender de más catálogos en este test.
        OrdenCompra.objects.create(proveedor=prov, monto_estimado=950, solicitud=solicitud_insumo)

    def test_faltantes_no_suma_ledgers_internos_cuando_point_no_esta_disponible(self):
        unidad = UnidadMedida.objects.create(
            codigo="kg-faltantes-multi",
            nombre="Kilogramo Faltantes Multi",
            tipo=UnidadMedida.TIPO_MASA,
        )
        insumo = Insumo.objects.create(
            nombre="Insumo Faltantes Multi",
            codigo_point="FALTANTES-MULTI-001",
            unidad_base=unidad,
            activo=True,
        )
        ExistenciaInsumo.objects.create(
            insumo=insumo,
            almacen="ARMADO",
            stock_actual=Decimal("4"),
            punto_reorden=Decimal("100"),
        )
        ExistenciaInsumo.objects.create(
            insumo=insumo,
            almacen="ALMACEN_1",
            stock_actual=Decimal("2"),
            punto_reorden=Decimal("5"),
        )

        rows, _, _ = _faltantes_rows("all")
        row = next(item for item in rows if item.insumo.id == insumo.id)

        self.assertIsNone(row.stock_actual)
        self.assertEqual(row.punto_reorden, Decimal("5"))
        self.assertEqual(row.nivel, "unavailable")

    def test_bi_inventory_snapshot_agrupa_stock_y_usa_umbral_de_almacen_1(self):
        unidad = UnidadMedida.objects.create(
            codigo="kg-bi-inv-multi",
            nombre="Kilogramo BI Inventory Multi",
            tipo=UnidadMedida.TIPO_MASA,
        )
        insumo = Insumo.objects.create(
            nombre="Insumo BI Inventory Multi",
            unidad_base=unidad,
            activo=True,
        )
        ExistenciaInsumo.objects.create(
            insumo=insumo,
            almacen="ALMACEN_1",
            stock_actual=Decimal("2"),
            punto_reorden=Decimal("5"),
        )
        ExistenciaInsumo.objects.create(
            insumo=insumo,
            almacen="ARMADO",
            stock_actual=Decimal("4"),
            punto_reorden=Decimal("100"),
        )

        snapshot = _bi_inventory_snapshot()

        self.assertEqual(snapshot["total"], 1)
        self.assertEqual(snapshot["criticos"], 0)
        self.assertEqual(snapshot["bajo_reorden"], 0)
        self.assertEqual(snapshot["unavailable"], 1)

    def test_bi_view_renders(self):
        sucursal = self._create_sucursal("BI-01", "Sucursal BI 01")
        receta = Receta.objects.create(nombre="Pastel BI Histórico", hash_contenido="hash-bi-historico-001")
        fecha_actual = timezone.localdate() - timedelta(days=1)
        fecha_comparable = fecha_actual - timedelta(days=7)
        point_branch = PointBranch.objects.create(external_id="BI-01", name="Sucursal BI 01", erp_branch=sucursal)
        point_product = PointProduct.objects.create(external_id="PBI-01", sku="BI001", name="Pastel BI Histórico", active=True)
        VentaHistorica.objects.create(
            receta=receta,
            sucursal=sucursal,
            fecha=fecha_comparable,
            cantidad=Decimal("8"),
            monto_total=Decimal("800"),
            fuente="POINT_HIST_2026_Q1",
        )
        VentaHistorica.objects.create(
            receta=receta,
            sucursal=sucursal,
            fecha=fecha_actual,
            cantidad=Decimal("10"),
            monto_total=Decimal("1000"),
            fuente="POINT_HIST_2026_Q1",
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=receta,
            sale_date=fecha_actual,
            quantity=Decimal("10"),
            total_amount=Decimal("1000"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=receta,
            sale_date=fecha_comparable,
            quantity=Decimal("8"),
            total_amount=Decimal("800"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch,
            indicator_date=fecha_actual,
            total_amount=Decimal("1000"),
            total_tickets=4,
            total_avg_ticket=Decimal("250"),
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch,
            indicator_date=fecha_comparable,
            total_amount=Decimal("800"),
            total_tickets=4,
            total_avg_ticket=Decimal("200"),
        )
        resp = self.client.get(reverse("reportes:bi"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "BI Ejecutivo")
        self.assertContains(resp, "Tablero ejecutivo del negocio")
        self.assertContains(resp, "Tendencia semanal de venta")
        self.assertContains(resp, "Mes contra mismo mes del año anterior")
        self.assertContains(resp, "Margen vs volumen por producto")
        self.assertContains(resp, "Rendimiento por sucursal")
        self.assertContains(resp, "Producido contra vendido")
        self.assertContains(resp, "Flujo histórico mensual del centro")
        self.assertContains(resp, "Ticket promedio")
        self.assertContains(resp, "$1,000.00")
        self.assertContains(resp, "$250.00")
        self.assertContains(resp, "forecastTrendChart")
        self.assertContains(resp, "yoyMonthlyChart")
        self.assertContains(resp, "profitabilityScatterChart")
        self.assertContains(resp, "productionWeeklyChart")
        self.assertContains(resp, "productionCategoryChart")
        self.assertContains(resp, "centralFlowChart")
        self.assertNotContains(resp, "Ver control ERP del BI")
        self.assertNotContains(resp, "Cadena documental ERP")
        self.assertNotContains(resp, "Entrega de reportes a downstream")
        self.assertNotContains(resp, "Ruta troncal ERP")
        self.assertNotContains(resp, "Mesa de gobierno ERP")
        self.assertNotContains(resp, "Ruta crítica ERP")
        self.assertNotContains(resp, "Radar ejecutivo ERP")
        self.assertNotContains(resp, "Centro de mando ERP")
        self.assertNotContains(resp, "Madurez ERP de reportes")
        self.assertNotContains(resp, "Criterios de cierre ERP")
        self.assertIn("maturity_summary", resp.context)
        self.assertIn("daily_decision_rows", resp.context)
        self.assertIn("branch_weekday_rows", resp.context)
        self.assertIn("product_weekday_rows", resp.context)
        self.assertIn("handoff_map", resp.context)
        self.assertIn("release_gate_rows", resp.context)
        self.assertIn("release_gate_completion", resp.context)
        self.assertIn("dependency_status", resp.context["enterprise_chain"][0])
        self.assertIn("owner", resp.context["document_stage_rows"][0])
        self.assertIn("completion", resp.context["document_stage_rows"][0])
        self.assertIn("erp_governance_rows", resp.context)
        self.assertIn("critical_path_rows", resp.context)
        self.assertIn("executive_radar_rows", resp.context)
        self.assertIn("erp_command_center", resp.context)
        self.assertIn("ventas_historicas_summary", resp.context)
        self.assertIn("forecast_panel", resp.context)
        self.assertIn("yoy_panel", resp.context)
        self.assertIn("profitability_panel", resp.context)
        self.assertIn("branch_contribution_panel", resp.context)
        self.assertIn("budget_operating_panel", resp.context)
        self.assertIn("production_sales_panel", resp.context)
        self.assertIn("central_flow_panel", resp.context)
        self.assertIn("inventory_ledger_panel", resp.context)
        self.assertEqual(resp.context["ventas_historicas_summary"]["total_rows"], 2)

    @patch("reportes.executive_panels._operational_projection_for_week", return_value=None)
    @patch("reportes.executive_panels._indicator_daily_ticket_map")
    @patch("reportes.executive_panels._sales_fact_daily_map")
    def test_sales_forecast_panel_applies_minimum_high_scenario_buffer_when_history_is_stable(
        self,
        sales_daily_map_mock,
        ticket_daily_map_mock,
        _operational_projection_mock,
    ):
        latest_date = date(2026, 4, 12)
        current_week_start = latest_date - timedelta(days=latest_date.weekday())
        sales_daily_map = {}
        ticket_daily_map = {}
        for week_offset in range(8):
            week_start = current_week_start - timedelta(days=7 * week_offset)
            for day_offset in range(7):
                current_day = week_start + timedelta(days=day_offset)
                sales_daily_map[current_day] = (Decimal("100"), Decimal("10"))
                ticket_daily_map[current_day] = 1
        sales_daily_map_mock.return_value = sales_daily_map
        ticket_daily_map_mock.return_value = ticket_daily_map

        panel = build_sales_forecast_panel(latest_date=latest_date)

        self.assertEqual(panel["forecast_amount"], Decimal("700.00"))
        self.assertEqual(panel["high_scenario_amount"], Decimal("735.00"))
        self.assertEqual(panel["forecast_quantity"], Decimal("70.00"))
        self.assertEqual(panel["high_scenario_quantity"], Decimal("73.50"))
        self.assertFalse(panel["has_atypical_history"])
        self.assertIn("colchón mínimo de 5.00%", panel["basis_note"])
        self.assertIn("No se detectaron picos atípicos recientes", panel["high_scenario_note"])

    @patch("reportes.executive_panels._operational_projection_for_week", return_value=None)
    @patch("reportes.executive_panels._indicator_daily_ticket_map")
    @patch("reportes.executive_panels._sales_fact_daily_map")
    def test_sales_forecast_panel_uses_atypical_peak_for_high_scenario(
        self,
        sales_daily_map_mock,
        ticket_daily_map_mock,
        _operational_projection_mock,
    ):
        latest_date = date(2026, 4, 12)
        current_week_start = latest_date - timedelta(days=latest_date.weekday())
        sales_daily_map = {}
        ticket_daily_map = {}
        for week_offset in range(8):
            week_start = current_week_start - timedelta(days=7 * week_offset)
            daily_amount = Decimal("100")
            daily_quantity = Decimal("10")
            if week_offset == 5:
                daily_amount = Decimal("200")
                daily_quantity = Decimal("20")
            for day_offset in range(7):
                current_day = week_start + timedelta(days=day_offset)
                sales_daily_map[current_day] = (daily_amount, daily_quantity)
                ticket_daily_map[current_day] = 1
        sales_daily_map_mock.return_value = sales_daily_map
        ticket_daily_map_mock.return_value = ticket_daily_map

        panel = build_sales_forecast_panel(latest_date=latest_date)

        self.assertTrue(panel["has_atypical_history"])
        self.assertEqual(panel["forecast_amount"], Decimal("700.00"))
        self.assertEqual(panel["high_scenario_amount"], Decimal("1400.00"))
        self.assertEqual(panel["high_scenario_quantity"], Decimal("140.00"))
        self.assertIn("mayor pico atípico reciente", panel["high_scenario_note"])
        self.assertEqual(panel["atypical_rows"][0]["amount"], Decimal("1400.00"))

    @patch("reportes.executive_panels._operational_projection_for_week", return_value=None)
    @patch("reportes.executive_panels._indicator_daily_ticket_map")
    @patch("reportes.executive_panels._sales_fact_daily_map")
    def test_sales_forecast_panel_excludes_open_current_week_from_baseline(
        self,
        sales_daily_map_mock,
        ticket_daily_map_mock,
        _operational_projection_mock,
    ):
        latest_date = date(2026, 6, 23)
        current_week_start = latest_date - timedelta(days=latest_date.weekday())
        sales_daily_map = {}
        ticket_daily_map = {}
        weekly_amounts = {
            0: Decimal("100"),
            1: Decimal("600"),
            2: Decimal("620"),
            3: Decimal("640"),
        }
        for week_offset in range(8):
            week_start = current_week_start - timedelta(days=7 * week_offset)
            weekly_amount = weekly_amounts.get(week_offset, Decimal("650"))
            daily_amount = weekly_amount / Decimal("7")
            daily_quantity = Decimal("10")
            for day_offset in range(7):
                current_day = week_start + timedelta(days=day_offset)
                sales_daily_map[current_day] = (daily_amount, daily_quantity)
                ticket_daily_map[current_day] = 1
        sales_daily_map_mock.return_value = sales_daily_map
        ticket_daily_map_mock.return_value = ticket_daily_map

        panel = build_sales_forecast_panel(latest_date=latest_date)

        self.assertEqual(panel["forecast_amount"], Decimal("620.00"))
        self.assertEqual(panel["baseline_label"], "3 semana(s) base")

    @patch(
        "reportes.executive_panels._operational_projection_for_week",
        return_value={
            "amount": Decimal("900.00"),
            "quantity": Decimal("90.00"),
            "high_amount": Decimal("990.00"),
            "high_quantity": Decimal("99.00"),
            "method": "forecast-operativo-3-semanas",
            "products": 12,
        },
    )
    @patch("reportes.executive_panels._indicator_daily_ticket_map")
    @patch("reportes.executive_panels._sales_fact_daily_map")
    def test_sales_forecast_panel_prefers_operational_projection_when_available(
        self,
        sales_daily_map_mock,
        ticket_daily_map_mock,
        operational_projection_mock,
    ):
        latest_date = date(2026, 4, 12)
        current_week_start = latest_date - timedelta(days=latest_date.weekday())
        sales_daily_map = {}
        ticket_daily_map = {}
        for week_offset in range(8):
            week_start = current_week_start - timedelta(days=7 * week_offset)
            for day_offset in range(7):
                current_day = week_start + timedelta(days=day_offset)
                sales_daily_map[current_day] = (Decimal("100"), Decimal("10"))
                ticket_daily_map[current_day] = 1
        sales_daily_map_mock.return_value = sales_daily_map
        ticket_daily_map_mock.return_value = ticket_daily_map

        panel = build_sales_forecast_panel(latest_date=latest_date)

        self.assertEqual(panel["forecast_amount"], Decimal("900.00"))
        self.assertEqual(panel["forecast_quantity"], Decimal("90.00"))
        self.assertEqual(panel["baseline_label"], "Proyección operativa")
        self.assertEqual(panel["forecast_source"], "forecast-operativo-3-semanas")
        operational_projection_mock.assert_called_once_with(
            start_date=date(2026, 4, 13),
            end_date=date(2026, 4, 19),
        )

    def test_daily_snapshot_marks_missing_required_branch(self):
        fecha_actual = timezone.localdate() - timedelta(days=1)
        matriz = self._create_sucursal("MATRIZ", "Matriz")
        guamuchil = self._create_sucursal("GUAMUCHIL", "Guamuchil")
        guamuchil.fecha_apertura = fecha_actual
        guamuchil.save(update_fields=["fecha_apertura"])

        point_branch_matriz = PointBranch.objects.create(
            external_id="1",
            name="Matriz",
            status=PointBranch.STATUS_ACTIVE,
            erp_branch=matriz,
        )
        PointBranch.objects.create(
            external_id="13",
            name="Guamuchil",
            status=PointBranch.STATUS_ACTIVE,
            erp_branch=guamuchil,
        )
        point_product = PointProduct.objects.create(
            external_id="PBI-MISS-01",
            sku="MISS01",
            name="Pastel Corte Incompleto",
            active=True,
        )
        PointDailySale.objects.create(
            branch=point_branch_matriz,
            product=point_product,
            sale_date=fecha_actual,
            quantity=Decimal("5"),
            total_amount=Decimal("1000"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_matriz,
            indicator_date=fecha_actual,
            total_amount=Decimal("1000"),
            total_tickets=5,
            total_avg_ticket=Decimal("200"),
        )

        resp = self.client.get(reverse("reportes:ventas"))

        self.assertEqual(resp.status_code, 200)
        snapshot = resp.context["daily_sales_snapshot"]
        self.assertEqual(snapshot["status"], "Corte incompleto")
        self.assertEqual(snapshot["required_branch_count"], 2)
        self.assertEqual(snapshot["present_required_branch_count"], 1)
        self.assertEqual(snapshot["missing_required_branch_count"], 1)
        self.assertIn("Guamuchil", snapshot["missing_required_branch_names"])
        self.assertContains(resp, "Corte incompleto")
        self.assertContains(resp, "Guamuchil")

    def test_daily_snapshot_ignores_matrizdbg_alias_branch(self):
        fecha_actual = timezone.localdate() - timedelta(days=1)
        matriz = self._create_sucursal("MATRIZ", "Matriz")
        matriz_dbg = self._create_sucursal("MATRIZDBG", "Matriz DBG")
        point_branch_matriz = PointBranch.objects.create(
            external_id="1",
            name="Matriz",
            status=PointBranch.STATUS_ACTIVE,
            erp_branch=matriz,
        )
        PointBranch.objects.create(
            external_id="dbg1",
            name="Matriz DBG",
            status=PointBranch.STATUS_ACTIVE,
            erp_branch=matriz_dbg,
        )
        point_product = PointProduct.objects.create(
            external_id="PBI-MTZ-01",
            sku="MTZ01",
            name="Pastel Alias Matriz",
            active=True,
        )
        PointDailySale.objects.create(
            branch=point_branch_matriz,
            product=point_product,
            sale_date=fecha_actual,
            quantity=Decimal("5"),
            total_amount=Decimal("1000"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_matriz,
            indicator_date=fecha_actual,
            total_amount=Decimal("1000"),
            total_tickets=5,
            total_avg_ticket=Decimal("200"),
        )

        resp = self.client.get(reverse("reportes:ventas"))

        self.assertEqual(resp.status_code, 200)
        snapshot = resp.context["daily_sales_snapshot"]
        self.assertEqual(snapshot["required_branch_count"], 1)
        self.assertEqual(snapshot["present_required_branch_count"], 1)
        self.assertEqual(snapshot["missing_required_branch_count"], 0)
        self.assertNotIn("Matriz DBG", snapshot["missing_required_branch_names"])

    def test_daily_snapshot_counts_indicator_zero_amount_as_present_branch(self):
        fecha_actual = timezone.localdate() - timedelta(days=1)
        matriz = self._create_sucursal("MATRIZ", "Matriz")
        leyva = self._create_sucursal("LEYVA", "Leyva")
        point_branch_matriz = PointBranch.objects.create(
            external_id="1",
            name="Matriz",
            status=PointBranch.STATUS_ACTIVE,
            erp_branch=matriz,
        )
        point_branch_leyva = PointBranch.objects.create(
            external_id="2",
            name="Leyva",
            status=PointBranch.STATUS_ACTIVE,
            erp_branch=leyva,
        )
        point_product = PointProduct.objects.create(
            external_id="PBI-ZERO-01",
            sku="ZERO01",
            name="Pastel Cero Válido",
            active=True,
        )
        PointDailySale.objects.create(
            branch=point_branch_matriz,
            product=point_product,
            sale_date=fecha_actual,
            quantity=Decimal("5"),
            total_amount=Decimal("1000"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_matriz,
            indicator_date=fecha_actual,
            total_amount=Decimal("1000"),
            total_tickets=5,
            total_avg_ticket=Decimal("200"),
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_leyva,
            indicator_date=fecha_actual,
            total_amount=Decimal("0"),
            total_tickets=0,
            total_avg_ticket=Decimal("0"),
        )

        resp = self.client.get(reverse("reportes:ventas"))

        self.assertEqual(resp.status_code, 200)
        snapshot = resp.context["daily_sales_snapshot"]
        self.assertEqual(snapshot["required_branch_count"], 2)
        self.assertEqual(snapshot["present_required_branch_count"], 2)
        self.assertEqual(snapshot["missing_required_branch_count"], 0)
        self.assertNotIn("Leyva", snapshot["missing_required_branch_names"])

    def test_daily_snapshot_ignores_validated_operational_zero_exception(self):
        fecha_actual = date(2026, 4, 4)
        matriz = self._create_sucursal("MATRIZ", "Matriz")
        colosio = self._create_sucursal("COLOSIO", "Colosio")
        point_branch_colosio = PointBranch.objects.create(
            external_id="5",
            name="Colosio",
            status=PointBranch.STATUS_ACTIVE,
            erp_branch=colosio,
        )
        PointBranch.objects.create(
            external_id="1",
            name="Matriz",
            status=PointBranch.STATUS_ACTIVE,
            erp_branch=matriz,
        )
        point_product = PointProduct.objects.create(
            external_id="PBI-EXC-01",
            sku="EXC01",
            name="Pastel Excepción Operativa",
            active=True,
        )
        PointDailySale.objects.create(
            branch=point_branch_colosio,
            product=point_product,
            sale_date=fecha_actual,
            quantity=Decimal("5"),
            total_amount=Decimal("1000"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_colosio,
            indicator_date=fecha_actual,
            total_amount=Decimal("1000"),
            total_tickets=5,
            total_avg_ticket=Decimal("200"),
        )

        with patch("reportes.dashboard_sales_dataset.timezone.localdate", return_value=fecha_actual), patch(
            "reportes.views.timezone.localdate", return_value=fecha_actual
        ):
            snapshot = _bi_daily_sales_snapshot()

        self.assertEqual(snapshot["required_branch_count"], 2)
        self.assertEqual(snapshot["present_required_branch_count"], 2)
        self.assertEqual(snapshot["missing_required_branch_count"], 0)
        self.assertNotIn("Matriz", snapshot["missing_required_branch_names"])

    def test_daily_snapshot_counts_successful_official_backfill_zero_row_as_present_branch(self):
        fecha_actual = date(2026, 1, 4)
        matriz = self._create_sucursal("MATRIZ", "Matriz")
        plaza_nio = self._create_sucursal("PLAZA_NIO", "Plaza Nío")
        point_branch_matriz = PointBranch.objects.create(
            external_id="1",
            name="Matriz",
            status=PointBranch.STATUS_ACTIVE,
            erp_branch=matriz,
        )
        PointBranch.objects.create(
            external_id="8",
            name="Plaza Nío",
            status=PointBranch.STATUS_ACTIVE,
            erp_branch=plaza_nio,
        )
        point_product = PointProduct.objects.create(
            external_id="PBI-BACKFILL-ZERO-01",
            sku="BACKFILLZERO01",
            name="Pastel Backfill Zero",
            active=True,
        )
        PointDailySale.objects.create(
            branch=point_branch_matriz,
            product=point_product,
            sale_date=fecha_actual,
            quantity=Decimal("5"),
            total_amount=Decimal("1000"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch_matriz,
            indicator_date=fecha_actual,
            total_amount=Decimal("1000"),
            total_tickets=5,
            total_avg_ticket=Decimal("200"),
        )
        sync_job = PointSyncJob.objects.create(
            job_type=PointSyncJob.JOB_TYPE_SALES,
            status=PointSyncJob.STATUS_SUCCESS,
            parameters={"source": "POINT_OFFICIAL_REPORT"},
            result_summary={},
        )
        PointExtractionLog.objects.create(
            sync_job=sync_job,
            level=PointExtractionLog.LEVEL_INFO,
            message="Backfill oficial 8 2026-01-04",
            context={
                "branch": "Plaza Nío",
                "branch_external_id": "8",
                "sale_date": fecha_actual.isoformat(),
                "rows_imported": 0,
                "rows_deleted": 0,
                "reports_downloaded": 1,
            },
        )

        with patch("reportes.dashboard_sales_dataset.timezone.localdate", return_value=fecha_actual), patch(
            "reportes.views.timezone.localdate", return_value=fecha_actual
        ):
            snapshot = _bi_daily_sales_snapshot()

        self.assertEqual(snapshot["required_branch_count"], 2)
        self.assertEqual(snapshot["present_required_branch_count"], 2)
        self.assertEqual(snapshot["missing_required_branch_count"], 0)
        self.assertNotIn("Plaza Nío", snapshot["missing_required_branch_names"])

    def test_daily_snapshot_prefers_official_cut_when_present(self):
        fecha_actual = timezone.localdate() - timedelta(days=1)
        matriz = self._create_sucursal("MATRIZ", "Matriz")
        point_branch = PointBranch.objects.create(
            external_id="1",
            name="Matriz",
            status=PointBranch.STATUS_ACTIVE,
            erp_branch=matriz,
        )
        point_product = PointProduct.objects.create(
            external_id="PBI-OFC-01",
            sku="OFC01",
            name="Pastel Corte Oficial",
            active=True,
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            sale_date=fecha_actual,
            quantity=Decimal("5"),
            total_amount=Decimal("1000"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch,
            indicator_date=fecha_actual,
            total_amount=Decimal("1000"),
            total_tickets=5,
            total_avg_ticket=Decimal("200"),
        )
        CorteOficialDiario.objects.create(
            corte_date=fecha_actual,
            total_amount=Decimal("138760.98"),
            total_tickets=473,
            avg_ticket=Decimal("293.36"),
            contado_amount=Decimal("138760.98"),
            credito_amount=Decimal("0"),
            discounts_amount=Decimal("3095.02"),
            new_customers=4,
            evidence_path="/tmp/corte_oficial_2026-04-01.png",
        )

        resp = self.client.get(reverse("reportes:ventas"))

        self.assertEqual(resp.status_code, 200)
        snapshot = resp.context["daily_sales_snapshot"]
        self.assertTrue(snapshot["official_cut_applied"])
        self.assertEqual(snapshot["total_amount"], Decimal("138760.98"))
        self.assertEqual(snapshot["total_tickets"], 473)
        self.assertEqual(snapshot["avg_ticket"], Decimal("293.36"))
        self.assertContains(resp, "Corte oficial aplicado")

    def test_bi_view_renders_budget_operating_panel(self):
        detail_admin = PresupuestoImport.objects.create(
            tipo=PresupuestoImport.TIPO_DETALLE,
            fuente_nombre="PRESUPUESTO ADMINISTRACIÓN 2026.xlsx",
            archivo_ruta="/tmp/admin-detalle.xlsx",
            archivo_hash="hash-admin-detalle-bi",
            sheet_name="ADMON",
            titulo="ADMON",
            metadata={"kind": "admin_recurrente"},
        )
        detail_sales = PresupuestoImport.objects.create(
            tipo=PresupuestoImport.TIPO_DETALLE,
            fuente_nombre="PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx",
            archivo_ruta="/tmp/sales-detalle.xlsx",
            archivo_hash="hash-sales-detalle-bi",
            sheet_name="MATRIZ",
            titulo="MATRIZ",
            metadata={"kind": "branch_sales"},
        )
        detail_payroll = PresupuestoImport.objects.create(
            tipo=PresupuestoImport.TIPO_DETALLE,
            fuente_nombre="PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx",
            archivo_ruta="/tmp/nomina-detalle.xlsx",
            archivo_hash="hash-nomina-detalle-bi",
            sheet_name="ADMINISTRACION",
            titulo="ADMINISTRACION",
            metadata={"kind": "payroll_area"},
        )
        PresupuestoResumenMensual.objects.create(
            period=date(2026, 1, 1),
            tipo=PresupuestoResumenMensual.TIPO_GLOBAL,
            fuente_nombre="",
            total_budget=Decimal("100000"),
            total_actual=Decimal("95000"),
            total_variance=Decimal("-5"),
            line_count=10,
            metadata={"global_mode": "master_source"},
        )
        PresupuestoResumenMensual.objects.create(
            period=date(2026, 1, 1),
            tipo=PresupuestoResumenMensual.TIPO_FUENTE,
            fuente_nombre="PRESUPUESTO ADMINISTRACIÓN 2026.xlsx",
            total_budget=Decimal("70000"),
            total_actual=Decimal("65000"),
            total_variance=Decimal("-7.14"),
            line_count=5,
        )
        PresupuestoResumenMensual.objects.create(
            period=date(2026, 1, 1),
            tipo=PresupuestoResumenMensual.TIPO_FUENTE,
            fuente_nombre="PRESUPUESTO DE GASTOS VENTAS 2026 AUTORIZADO.xlsx",
            total_budget=Decimal("20000"),
            total_actual=Decimal("18000"),
            total_variance=Decimal("-10"),
            line_count=3,
        )
        PresupuestoResumenMensual.objects.create(
            period=date(2026, 1, 1),
            tipo=PresupuestoResumenMensual.TIPO_FUENTE,
            fuente_nombre="PRESUPUESTO NOMINA 2026 AUTORIZADO.xlsx",
            total_budget=Decimal("10000"),
            total_actual=Decimal("12000"),
            total_variance=Decimal("20"),
            line_count=2,
        )
        import_obj = PresupuestoImport.objects.create(
            tipo=PresupuestoImport.TIPO_GENERAL,
            fuente_nombre="PRESUPUESTO ADMINISTRACIÓN 2026.xlsx",
            archivo_ruta="/tmp/admin.xlsx",
            archivo_hash="hash-admin-bi",
            sheet_name="GENERAL",
            titulo="PRESUPUESTO GENERAL POLLYANA'S DOLCE",
        )
        PresupuestoLineaMensual.objects.create(
            importacion=import_obj,
            external_key="bi-admin-1",
            period=date(2026, 1, 1),
            concept="INGRESOS",
            monthly_budget=Decimal("40000"),
            monthly_actual=Decimal("38000"),
        )
        PresupuestoLineaMensual.objects.create(
            importacion=import_obj,
            external_key="bi-admin-2",
            period=date(2026, 1, 1),
            concept="Sueldo",
            monthly_budget=Decimal("30000"),
            monthly_actual=Decimal("28000"),
        )
        PresupuestoLineaMensual.objects.create(
            importacion=detail_admin,
            external_key="bi-detail-admin-1",
            period=date(2026, 1, 1),
            concept="Sueldo",
            audit_status=PresupuestoLineaMensual.AUDIT_OK,
            monthly_budget=Decimal("30000"),
            monthly_actual=Decimal("28000"),
            metadata={"kind": "admin_recurrente"},
        )
        PresupuestoLineaMensual.objects.create(
            importacion=detail_sales,
            external_key="bi-detail-sales-1",
            period=date(2026, 1, 1),
            concept="Arrendamiento local",
            audit_status=PresupuestoLineaMensual.AUDIT_OK,
            monthly_budget=Decimal("20000"),
            monthly_actual=Decimal("18000"),
            metadata={"kind": "branch_sales"},
        )
        PresupuestoLineaMensual.objects.create(
            importacion=detail_payroll,
            external_key="bi-detail-payroll-1",
            period=date(2026, 1, 1),
            concept="SUELDO",
            audit_status=PresupuestoLineaMensual.AUDIT_OK,
            monthly_budget=Decimal("10000"),
            monthly_actual=Decimal("12000"),
            metadata={"kind": "payroll_area"},
        )
        EmpresaResultadoMensual.objects.create(
            periodo=date(2026, 1, 1),
            venta_total=Decimal("120000"),
            costo_fabricacion_total=Decimal("50000"),
            margen_bruto_total=Decimal("70000"),
            gasto_comercial_total=Decimal("12000"),
            contribucion_total=Decimal("58000"),
            gasto_corporativo_total=Decimal("18000"),
            utilidad_operativa_total=Decimal("40000"),
            metadata={
                "venta_costeada_total": "115000",
                "venta_no_receta_total": "5000",
                "venta_receta_sin_match_total": "0",
                "venta_sin_mapear_total": "5000",
            },
        )
        ProductoPricingDecisionMensual.objects.create(
            periodo=date(2026, 1, 1),
            receta=Receta.objects.create(nombre="Producto Pricing BI", hash_contenido="hash-pricing-bi-001"),
            asp_actual=Decimal("10"),
            costo_fabricacion_unit=Decimal("5"),
            contribucion_unit=Decimal("2"),
            margen_bruto_pct=Decimal("50"),
            margen_contribucion_pct=Decimal("20"),
            precio_objetivo_bruto=Decimal("11"),
            precio_objetivo_contribucion=Decimal("12"),
            gap_precio=Decimal("1"),
            accion_sugerida="SUBIR_PRECIO",
            impacto_estimado=Decimal("1000"),
            metadata={},
        )

        resp = self.client.get(reverse("reportes:bi"))

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Meta de venta, costo, gasto y resultado")
        self.assertContains(resp, "Resultado 2026")
        self.assertContains(resp, "Meta de venta YTD")
        self.assertContains(resp, "Costo producto YTD")
        self.assertContains(resp, "Gasto recurrente YTD")
        self.assertContains(resp, "Resultado YTD")
        self.assertContains(resp, "Costo + gasto verificado")
        self.assertContains(resp, "Administración recurrente")
        self.assertContains(resp, "Comercial sucursales")
        self.assertContains(resp, "Nómina por área")
        self.assertContains(resp, "$30,000.00")
        self.assertContains(resp, "$28,000.00")
        self.assertContains(resp, "Presupuesto maestro")
        self.assertContains(resp, "Presupuesto por sucursal")
        self.assertContains(resp, "$100,000.00")
        self.assertContains(resp, "$120,000.00")
        self.assertContains(resp, "$40,000.00")
        self.assertContains(resp, "$60,000.00")
        self.assertContains(resp, "$-20,000.00")

    def test_bi_budget_panel_prefers_monthly_official_sales_total(self):
        EmpresaResultadoMensual.objects.create(
            periodo=date(2026, 3, 1),
            venta_total=Decimal("3449037.19"),
            costo_fabricacion_total=Decimal("1197704.04"),
            margen_bruto_total=Decimal("2251333.15"),
            gasto_comercial_total=Decimal("0"),
            contribucion_total=Decimal("2251333.15"),
            gasto_corporativo_total=Decimal("0"),
            utilidad_operativa_total=Decimal("1925162.15"),
            metadata={"venta_costeada_total": "0", "venta_no_receta_total": "0"},
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            total_quantity=Decimal("0"),
            gross_amount=Decimal("3326094.25"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("3326094.25"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("3326094.25"),
            raw_payload={"source": "direccion_general"},
        )

        resp = self.client.get(reverse("reportes:bi"), {"budget_month": "3"})

        self.assertEqual(resp.status_code, 200)
        panel = resp.context["budget_operating_panel"]
        self.assertEqual(panel["selected_row"]["sales_total"], Decimal("3326094.25"))
        self.assertFalse(panel["selected_row"]["finance_close_complete"])
        self.assertEqual(panel["selected_row"]["health_signal"]["label"], "Lectura parcial")
        self.assertEqual(panel["selected_row"]["gross_margin_signal"]["label"], "Sano")
        self.assertEqual(panel["selected_row"]["gross_margin_total"], Decimal("2128390.21"))
        self.assertContains(resp, "Lectura parcial")
        self.assertContains(resp, "Cierre pendiente")
        self.assertContains(resp, "Margen bruto real")
        self.assertContains(resp, "Lectura costo vs precio global")
        self.assertContains(resp, "costo de producción pesa")
        self.assertContains(resp, "64.0%")
        self.assertNotContains(resp, "Cierre completo")

    def test_bi_view_renders_branch_contribution_panel(self):
        sucursal = self._create_sucursal("SUC-BI", "Sucursal BI Rendimiento")
        receta = Receta.objects.create(nombre="Producto Sucursal BI", hash_contenido="hash-sucursal-bi-001")
        PointBranch.objects.create(external_id="PB-SUC-BI", name="Sucursal BI Rendimiento", erp_branch=sucursal)
        ProductoSucursalContribucionMensual.objects.create(
            periodo=date(2026, 2, 1),
            receta=receta,
            sucursal=sucursal,
            unidades_vendidas=Decimal("80"),
            venta_total=Decimal("16000"),
            asp=Decimal("200"),
            costo_producto_unit=Decimal("90"),
            costo_producto_total=Decimal("7200"),
            gasto_comercial_unit=Decimal("20"),
            gasto_comercial_total=Decimal("1600"),
            contribucion_total=Decimal("7200"),
            contribucion_unit=Decimal("90"),
            margen_contribucion_pct=Decimal("45"),
            metadata={},
        )
        ProductoSucursalContribucionMensual.objects.create(
            periodo=date(2026, 3, 1),
            receta=receta,
            sucursal=sucursal,
            unidades_vendidas=Decimal("100"),
            venta_total=Decimal("25000"),
            asp=Decimal("250"),
            costo_producto_unit=Decimal("100"),
            costo_producto_total=Decimal("10000"),
            gasto_comercial_unit=Decimal("25"),
            gasto_comercial_total=Decimal("2500"),
            contribucion_total=Decimal("12500"),
            contribucion_unit=Decimal("125"),
            margen_contribucion_pct=Decimal("50"),
            metadata={},
        )
        point_branch = PointBranch.objects.get(external_id="PB-SUC-BI")
        point_product = PointProduct.objects.create(external_id="NP-SVC-01", sku="SVC01", name="Servicio de domicilio", active=True)
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=None,
            sale_date=date(2026, 3, 20),
            quantity=Decimal("2"),
            total_amount=Decimal("350"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        point_product_resale = PointProduct.objects.create(
            external_id="NP-TE-01",
            sku="TE01",
            name="Te Chai",
            category="Te",
            active=True,
            metadata={"family": "Bebidas"},
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product_resale,
            receta=None,
            sale_date=date(2026, 3, 21),
            quantity=Decimal("3"),
            total_amount=Decimal("180"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )

        resp = self.client.get(reverse("reportes:bi"))

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Rendimiento por sucursal")
        self.assertContains(resp, "Sucursal BI Rendimiento")
        self.assertContains(resp, "$25,000.00")
        self.assertContains(resp, "$12,500.00")
        self.assertContains(resp, "Vs prom. YTD venta")
        self.assertContains(resp, "Reventa residual")
        self.assertContains(resp, "Servicio")
        self.assertContains(resp, "$180.00")

    def test_branch_contribution_panel_groups_non_recipe_sales_before_classifying(self):
        sucursal = self._create_sucursal("SUC-NR", "Sucursal No Receta")
        point_branch = PointBranch.objects.create(external_id="PB-SUC-NR", name="Sucursal No Receta", erp_branch=sucursal)
        point_product = PointProduct.objects.create(
            external_id="NR-SVC-01",
            sku="SVC01",
            name="Servicio de domicilio",
            active=True,
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=None,
            sale_date=date(2026, 3, 20),
            quantity=Decimal("2"),
            total_amount=Decimal("350"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=None,
            sale_date=date(2026, 3, 21),
            quantity=Decimal("1"),
            total_amount=Decimal("175"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )

        with patch.object(PointSalesMatchingService, "infer_cost_mode", wraps=PointSalesMatchingService().infer_cost_mode) as infer_cost_mode:
            panel = build_branch_contribution_panel(year=2026)

        self.assertEqual(infer_cost_mode.call_count, 1)
        row = panel["rows"][0]
        self.assertEqual(row["non_recipe_total"], Decimal("525"))
        self.assertEqual(row["non_recipe_service_total"], Decimal("525"))

    def test_bi_view_renders_branch_pricing_panel_for_selected_branch(self):
        sucursal = self._create_sucursal("SUC-PRICE", "Sucursal Pricing")
        receta_a = Receta.objects.create(nombre="Producto Defender", hash_contenido="hash-branch-price-001")
        receta_b = Receta.objects.create(nombre="Producto Subir", hash_contenido="hash-branch-price-002")
        ProductoSucursalContribucionMensual.objects.create(
            periodo=date(2026, 3, 1),
            receta=receta_a,
            sucursal=sucursal,
            unidades_vendidas=Decimal("100"),
            venta_total=Decimal("30000"),
            asp=Decimal("300"),
            costo_producto_unit=Decimal("120"),
            costo_producto_total=Decimal("12000"),
            gasto_comercial_unit=Decimal("30"),
            gasto_comercial_total=Decimal("3000"),
            contribucion_total=Decimal("15000"),
            contribucion_unit=Decimal("150"),
            margen_contribucion_pct=Decimal("50"),
            metadata={},
        )
        ProductoSucursalContribucionMensual.objects.create(
            periodo=date(2026, 3, 1),
            receta=receta_b,
            sucursal=sucursal,
            unidades_vendidas=Decimal("80"),
            venta_total=Decimal("12000"),
            asp=Decimal("150"),
            costo_producto_unit=Decimal("100"),
            costo_producto_total=Decimal("8000"),
            gasto_comercial_unit=Decimal("25"),
            gasto_comercial_total=Decimal("2000"),
            contribucion_total=Decimal("2000"),
            contribucion_unit=Decimal("25"),
            margen_contribucion_pct=Decimal("16.66"),
            metadata={},
        )

        resp = self.client.get(reverse("reportes:bi"), {"branch_id": sucursal.id})

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Pricing por tienda")
        self.assertContains(resp, "Sucursal Pricing")
        self.assertContains(resp, "Producto Defender")
        self.assertContains(resp, "Producto Subir")
        self.assertIn("branch_pricing_panel", resp.context)
        self.assertEqual(resp.context["branch_pricing_panel"]["selected_branch_id"], sucursal.id)

        resp_filtered = self.client.get(reverse("reportes:bi"), {"branch_id": sucursal.id, "action": "Corregir costo"})
        self.assertEqual(resp_filtered.status_code, 200)
        filtered_labels = [row["label"] for row in resp_filtered.context["branch_pricing_panel"]["rows"]]
        self.assertIn("Producto Subir", filtered_labels)
        self.assertNotIn("Producto Defender", filtered_labels)
        self.assertEqual(resp_filtered.context["branch_pricing_panel"]["selected_action"], "Corregir costo")

    def test_bi_branch_exports_use_selected_branch_panel(self):
        sucursal = self._create_sucursal("SUC-EXP", "Sucursal Export")
        receta = Receta.objects.create(nombre="Producto Export", hash_contenido="hash-branch-export-001")
        ProductoSucursalContribucionMensual.objects.create(
            periodo=date(2026, 3, 1),
            receta=receta,
            sucursal=sucursal,
            unidades_vendidas=Decimal("50"),
            venta_total=Decimal("10000"),
            asp=Decimal("200"),
            costo_producto_unit=Decimal("90"),
            costo_producto_total=Decimal("4500"),
            gasto_comercial_unit=Decimal("20"),
            gasto_comercial_total=Decimal("1000"),
            contribucion_total=Decimal("4500"),
            contribucion_unit=Decimal("90"),
            margen_contribucion_pct=Decimal("45"),
            metadata={},
        )

        resp_csv = self.client.get(reverse("reportes:bi"), {"branch_id": sucursal.id, "export": "csv"})
        self.assertEqual(resp_csv.status_code, 200)
        self.assertIn("bi_sucursal_suc-exp.csv", resp_csv["Content-Disposition"])
        self.assertContains(resp_csv, "Sucursal Export")
        self.assertContains(resp_csv, "Producto Export")

        resp_xlsx = self.client.get(reverse("reportes:bi"), {"branch_id": sucursal.id, "export": "xlsx"})
        self.assertEqual(resp_xlsx.status_code, 200)
        self.assertIn("bi_sucursal_suc-exp.xlsx", resp_xlsx["Content-Disposition"])

        resp_pdf = self.client.get(reverse("reportes:bi"), {"branch_id": sucursal.id, "export": "pdf"})
        self.assertEqual(resp_pdf.status_code, 200)
        self.assertEqual(resp_pdf["Content-Type"], "application/pdf")
        self.assertIn("bi_sucursal_suc-exp.pdf", resp_pdf["Content-Disposition"])

    def test_bi_branch_summary_exports_render_ytd_and_latest_month(self):
        sucursal = self._create_sucursal("SUC-SUM", "Sucursal Sumario")
        receta = Receta.objects.create(nombre="Producto Sumario", hash_contenido="hash-branch-summary-001")
        ProductoSucursalContribucionMensual.objects.create(
            periodo=date(2026, 2, 1),
            receta=receta,
            sucursal=sucursal,
            unidades_vendidas=Decimal("25"),
            venta_total=Decimal("5000"),
            asp=Decimal("200"),
            costo_producto_unit=Decimal("90"),
            costo_producto_total=Decimal("2250"),
            gasto_comercial_unit=Decimal("20"),
            gasto_comercial_total=Decimal("500"),
            contribucion_total=Decimal("2250"),
            contribucion_unit=Decimal("90"),
            margen_contribucion_pct=Decimal("45"),
            metadata={},
        )
        ProductoSucursalContribucionMensual.objects.create(
            periodo=date(2026, 3, 1),
            receta=receta,
            sucursal=sucursal,
            unidades_vendidas=Decimal("40"),
            venta_total=Decimal("10000"),
            asp=Decimal("250"),
            costo_producto_unit=Decimal("100"),
            costo_producto_total=Decimal("4000"),
            gasto_comercial_unit=Decimal("25"),
            gasto_comercial_total=Decimal("1000"),
            contribucion_total=Decimal("5000"),
            contribucion_unit=Decimal("125"),
            margen_contribucion_pct=Decimal("50"),
            metadata={},
        )

        resp_csv = self.client.get(reverse("reportes:bi"), {"export": "branches_csv"})
        self.assertEqual(resp_csv.status_code, 200)
        self.assertIn("bi_sucursales_2026.csv", resp_csv["Content-Disposition"])
        self.assertContains(resp_csv, "Sucursal Sumario")
        self.assertContains(resp_csv, "Venta ultimo mes")

        resp_xlsx = self.client.get(reverse("reportes:bi"), {"export": "branches_xlsx"})
        self.assertEqual(resp_xlsx.status_code, 200)
        self.assertIn("bi_sucursales_2026.xlsx", resp_xlsx["Content-Disposition"])

        resp_pdf = self.client.get(reverse("reportes:bi"), {"export": "branches_pdf"})
        self.assertEqual(resp_pdf.status_code, 200)
        self.assertEqual(resp_pdf["Content-Type"], "application/pdf")
        self.assertIn("bi_sucursales_2026.pdf", resp_pdf["Content-Disposition"])

        resp = self.client.get(reverse("reportes:bi"))
        self.assertContains(resp, "Venta $10,000.00")
        self.assertContains(resp, "semáforo")
        self.assertContains(resp, "Vs prom. YTD venta")
        self.assertContains(resp, "Brecha margen")
        self.assertContains(resp, "Último mes vs prom. YTD contribución")

    def test_ventas_view_renders(self):
        sucursal = self._create_sucursal("VTA-01", "Sucursal Ventas 01")
        receta = Receta.objects.create(nombre="Pastel Ventas", hash_contenido="hash-ventas-001")
        fecha_actual = timezone.localdate() - timedelta(days=1)
        fecha_comparable = fecha_actual - timedelta(days=7)
        point_branch = PointBranch.objects.create(external_id="VTA-01", name="Sucursal Ventas 01", erp_branch=sucursal)
        point_product = PointProduct.objects.create(external_id="PVTA-01", sku="VTA001", name="Pastel Ventas", active=True)
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=receta,
            sale_date=fecha_actual,
            quantity=Decimal("12"),
            total_amount=Decimal("1200"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=receta,
            sale_date=fecha_comparable,
            quantity=Decimal("10"),
            total_amount=Decimal("1000"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch,
            indicator_date=fecha_actual,
            total_amount=Decimal("1200"),
            total_tickets=5,
            total_avg_ticket=Decimal("240"),
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch,
            indicator_date=fecha_comparable,
            total_amount=Decimal("1000"),
            total_tickets=5,
            total_avg_ticket=Decimal("200"),
        )

        resp = self.client.get(reverse("reportes:ventas"))

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Panel comercial")
        self.assertContains(resp, "Tendencia semanal")
        self.assertContains(resp, "Mes contra año anterior")
        self.assertContains(resp, "Top sucursales")
        self.assertContains(resp, "Top productos")
        self.assertContains(resp, "Sucursales en alerta")
        self.assertContains(resp, "Productos en alerta")
        self.assertContains(resp, "Histórico comercial disponible")
        self.assertContains(resp, "salesForecastTrendChart")
        self.assertContains(resp, "salesYoyChart")
        self.assertContains(resp, "salesBranchChart")
        self.assertContains(resp, "salesProductChart")
        self.assertContains(resp, "salesBranchAlertChart")
        self.assertContains(resp, "salesProductAlertChart")
        self.assertContains(resp, "$1,200.00")
        self.assertContains(resp, "$240.00")
        self.assertTrue(resp.context["daily_sales_snapshot"]["top_branches"])
        self.assertIn("action_url", resp.context["daily_sales_snapshot"]["top_branches"][0])
        self.assertTrue(resp.context["branch_weekday_rows"])
        self.assertIn("action_url", resp.context["branch_weekday_rows"][0])
        self.assertTrue(resp.context["product_weekday_rows"])
        self.assertIn("action_url", resp.context["product_weekday_rows"][0])

    def test_branch_weekday_comparisons_uses_bulk_branch_rows_and_prefers_indicators(self):
        sucursal = self._create_sucursal("VTA-BULK", "Sucursal Bulk")
        fecha_actual = timezone.localdate() - timedelta(days=1)
        fecha_comparable = fecha_actual - timedelta(days=7)
        point_branch = PointBranch.objects.create(external_id="VTA-BULK", name="Sucursal Bulk", erp_branch=sucursal)
        point_product = PointProduct.objects.create(external_id="PVTA-BULK", sku="VTABULK", name="Pastel Bulk", active=True)

        for target_day, quantity, amount in (
            (fecha_actual, Decimal("10"), Decimal("1000")),
            (fecha_comparable, Decimal("7"), Decimal("700")),
        ):
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product,
                sale_date=target_day,
                quantity=quantity,
                total_amount=amount,
                source_endpoint="/Report/PrintReportes?idreporte=3",
            )

        for target_day, quantity, amount in (
            (fecha_actual, Decimal("12"), Decimal("1200")),
            (fecha_comparable, Decimal("9"), Decimal("850")),
        ):
            PointSalesDailyCategoryFact.objects.create(
                branch=point_branch,
                sale_date=target_day,
                sucursal_nombre=sucursal.nombre,
                categoria="Pasteles",
                total_cantidad=quantity,
                total_descuento=Decimal("0"),
                total_venta=amount,
                total_impuestos=Decimal("0"),
                total_venta_neta=amount,
            )

        PointDailyBranchIndicator.objects.create(
            branch=point_branch,
            indicator_date=fecha_actual,
            total_amount=Decimal("1300"),
            total_tickets=5,
            total_avg_ticket=Decimal("260"),
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch,
            indicator_date=fecha_comparable,
            total_amount=Decimal("900"),
            total_tickets=4,
            total_avg_ticket=Decimal("225"),
        )

        rows = _bi_branch_weekday_comparisons(limit=5)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["branch_code"], sucursal.codigo)
        self.assertEqual(rows[0]["branch_name"], sucursal.nombre)
        self.assertEqual(rows[0]["units"], Decimal("12"))
        self.assertEqual(rows[0]["amount"], Decimal("1300"))
        self.assertEqual(rows[0]["tickets"], 5)
        self.assertEqual(rows[0]["tone"], "success")
        self.assertEqual(rows[0]["status"], "Arriba del comparable")
        self.assertIn(fecha_comparable.isoformat(), rows[0]["detail"])
        self.assertIn(f"sucursal_id={sucursal.id}", rows[0]["action_url"])

    def test_product_weekday_comparisons_uses_bulk_product_rows_and_falls_back_to_product_name(self):
        sucursal = self._create_sucursal("VTA-PROD", "Sucursal Producto")
        fecha_actual = timezone.localdate() - timedelta(days=1)
        fecha_comparable = fecha_actual - timedelta(days=7)
        point_branch = PointBranch.objects.create(external_id="VTA-PROD", name="Sucursal Producto", erp_branch=sucursal)
        point_product = PointProduct.objects.create(external_id="PVTA-PROD", sku="VTAPROD", name="Servicio de domicilio", active=True)

        for target_day, quantity, amount in (
            (fecha_actual, Decimal("3"), Decimal("150")),
            (fecha_comparable, Decimal("2"), Decimal("80")),
        ):
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product,
                receta=None,
                sale_date=target_day,
                quantity=quantity,
                total_amount=amount,
                source_endpoint="/Report/PrintReportes?idreporte=3",
            )

        for target_day, quantity, amount in (
            (fecha_actual, Decimal("4"), Decimal("180")),
            (fecha_comparable, Decimal("2"), Decimal("90")),
        ):
            PointSalesDailyProductFact.objects.create(
                branch=point_branch,
                sale_date=target_day,
                sucursal_nombre=sucursal.nombre,
                categoria="Servicios",
                producto_nombre_historico="Servicio de domicilio",
                point_product=point_product,
                receta=None,
                total_cantidad=quantity,
                total_venta=amount,
            )

        rows = _bi_product_weekday_comparisons(limit=5)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["recipe_name"], "Servicio de domicilio")
        self.assertEqual(rows[0]["units"], Decimal("4"))
        self.assertEqual(rows[0]["amount"], Decimal("180"))
        self.assertEqual(rows[0]["branch_count"], 1)
        self.assertEqual(rows[0]["tone"], "success")
        self.assertEqual(rows[0]["status"], "Arriba del comparable")
        self.assertIn(fecha_comparable.isoformat(), rows[0]["detail"])
        self.assertIn("Servicio+de+domicilio", rows[0]["action_url"])

    def test_sales_source_context_detects_canonical_point_date_without_stage(self):
        sucursal = self._create_sucursal("CTX-V2", "Sucursal Contexto V2")
        point_branch = PointBranch.objects.create(external_id="CTX-V2", name="Sucursal Contexto V2", erp_branch=sucursal)
        latest_day = timezone.localdate() - timedelta(days=1)
        previous_day = latest_day - timedelta(days=7)

        PointSalesDailyCategoryFact.objects.create(
            branch=point_branch,
            sale_date=previous_day,
            sucursal_nombre=sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("5"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("500"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("500"),
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=point_branch,
            sale_date=latest_day,
            sucursal_nombre=sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("7"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("700"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("700"),
        )

        source = _sales_source_context()
        previous_dates = _sales_previous_dates(source, latest_day)

        self.assertEqual(source["mode"], "point_stage")
        self.assertEqual(source["latest_date"], latest_day)
        self.assertEqual(source["canonical_latest_date"], latest_day)
        self.assertIsNone(source["stage_latest_date"])
        self.assertEqual(previous_dates, [previous_day])

    def test_sales_source_context_keeps_stage_latest_date_when_stage_lags_canonical(self):
        sucursal = self._create_sucursal("CTX-STAGE", "Sucursal Contexto Stage")
        point_branch = PointBranch.objects.create(external_id="CTX-STAGE", name="Sucursal Contexto Stage", erp_branch=sucursal)
        stage_day = timezone.localdate() - timedelta(days=2)
        canonical_day = timezone.localdate() - timedelta(days=1)

        PointDailySale.objects.create(
            branch=point_branch,
            product=PointProduct.objects.create(external_id="CTX-STAGE-P", sku="CTXSTAGE", name="Producto Stage", active=True),
            sale_date=stage_day,
            quantity=Decimal("4"),
            total_amount=Decimal("400"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=point_branch,
            sale_date=canonical_day,
            sucursal_nombre=sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("8"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("800"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("800"),
        )

        source = _sales_source_context()

        self.assertEqual(source["mode"], "point_stage")
        self.assertEqual(source["latest_date"], stage_day)
        self.assertEqual(source["stage_latest_date"], stage_day)
        self.assertEqual(source["canonical_latest_date"], canonical_day)

    def test_ventas_historicas_bi_summary_uses_canonical_source_when_stage_is_missing(self):
        sucursal = self._create_sucursal("HIST-V2", "Sucursal Historico V2")
        point_branch = PointBranch.objects.create(external_id="HIST-V2", name="Sucursal Historico V2", erp_branch=sucursal)
        first_day = timezone.localdate() - timedelta(days=8)
        last_day = timezone.localdate() - timedelta(days=1)

        PointSalesDailyCategoryFact.objects.create(
            branch=point_branch,
            sale_date=first_day,
            sucursal_nombre=sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("5"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("500"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("500"),
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=point_branch,
            sale_date=last_day,
            sucursal_nombre=sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("7"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("700"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("700"),
        )
        PointSalesDailyProductFact.objects.create(
            branch=point_branch,
            sale_date=first_day,
            sucursal_nombre=sucursal.nombre,
            categoria="Pasteles",
            producto_nombre_historico="Pastel Historico V2",
            point_product=PointProduct.objects.create(external_id="HIST-V2-P", sku="HISTV2", name="Pastel Historico V2", active=True),
            receta=Receta.objects.create(nombre="Pastel Historico V2", hash_contenido="hist-v2-summary-001"),
            total_cantidad=Decimal("5"),
            total_venta=Decimal("500"),
        )

        summary = _ventas_historicas_bi_summary()

        self.assertTrue(summary["available"])
        self.assertEqual(summary["source_label"], "Point directo")
        self.assertEqual(summary["total_amount"], Decimal("1200"))
        self.assertEqual(summary["active_days"], 2)
        self.assertIn(first_day.strftime("%d/%m/%Y"), summary["date_label"])
        self.assertIn(last_day.strftime("%d/%m/%Y"), summary["date_label"])

    def test_ventas_historicas_bi_summary_mentions_canonical_date_when_stage_lags(self):
        sucursal = self._create_sucursal("HIST-LAG", "Sucursal Historico Lag")
        point_branch = PointBranch.objects.create(external_id="HIST-LAG", name="Sucursal Historico Lag", erp_branch=sucursal)
        stage_day = timezone.localdate() - timedelta(days=2)
        canonical_day = timezone.localdate() - timedelta(days=1)
        point_product = PointProduct.objects.create(external_id="HIST-LAG-P", sku="HISTLAG", name="Pastel Lag", active=True)

        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            sale_date=stage_day,
            quantity=Decimal("4"),
            total_amount=Decimal("400"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=point_branch,
            sale_date=canonical_day,
            sucursal_nombre=sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("8"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("800"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("800"),
        )

        summary = _ventas_historicas_bi_summary()

        self.assertTrue(summary["available"])
        self.assertEqual(summary["source_label"], "Point directo")
        self.assertIn(canonical_day.strftime("%d/%m/%Y"), summary["detail"])
        self.assertIn(stage_day.strftime("%d/%m/%Y"), summary["detail"])

    def test_bi_daily_sales_snapshot_uses_canonical_daily_totals_and_tops_without_stage(self):
        sucursal = self._create_sucursal("SNAP-V2", "Sucursal Snapshot V2")
        point_branch = PointBranch.objects.create(external_id="SNAP-V2", name="Sucursal Snapshot V2", erp_branch=sucursal)
        latest_day = timezone.localdate() - timedelta(days=1)
        point_product = PointProduct.objects.create(external_id="SNAP-V2-P", sku="SNAPV2", name="Pastel Snapshot", active=True)

        PointSalesDailyCategoryFact.objects.create(
            branch=point_branch,
            sale_date=latest_day,
            sucursal_nombre=sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("6"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("600"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("600"),
        )
        PointSalesDailyProductFact.objects.create(
            branch=point_branch,
            sale_date=latest_day,
            sucursal_nombre=sucursal.nombre,
            categoria="Pasteles",
            producto_nombre_historico="Pastel Snapshot",
            point_product=point_product,
            receta=Receta.objects.create(nombre="Pastel Snapshot", hash_contenido="snap-v2-001"),
            total_cantidad=Decimal("6"),
            total_venta=Decimal("600"),
        )
        PointDailyBranchIndicator.objects.create(
            branch=point_branch,
            indicator_date=latest_day,
            total_amount=Decimal("610"),
            total_tickets=3,
            total_avg_ticket=Decimal("203.33"),
        )

        snapshot = _bi_daily_sales_snapshot()

        self.assertEqual(snapshot["source_label"], "Point directo")
        self.assertEqual(snapshot["date_label"], latest_day.isoformat())
        self.assertEqual(snapshot["total_units"], Decimal("6"))
        self.assertEqual(snapshot["raw_total_amount"], Decimal("600"))
        self.assertEqual(snapshot["total_amount"], Decimal("610"))
        self.assertEqual(snapshot["total_tickets"], 3)
        self.assertEqual(snapshot["branch_count"], 1)
        self.assertEqual(snapshot["recipe_count"], 1)
        self.assertEqual(snapshot["top_branches"][0]["label"], sucursal.codigo)
        self.assertEqual(snapshot["top_products"][0]["label"], "Pastel Snapshot")

    def test_bi_daily_sales_snapshot_uses_canonical_previous_day_comparison_when_stage_missing(self):
        sucursal = self._create_sucursal("SNAP-COMP", "Sucursal Snapshot Compare")
        point_branch = PointBranch.objects.create(external_id="SNAP-COMP", name="Sucursal Snapshot Compare", erp_branch=sucursal)
        latest_day = timezone.localdate() - timedelta(days=1)
        prev_day = latest_day - timedelta(days=7)

        PointSalesDailyCategoryFact.objects.create(
            branch=point_branch,
            sale_date=prev_day,
            sucursal_nombre=sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("4"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("400"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("400"),
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=point_branch,
            sale_date=latest_day,
            sucursal_nombre=sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("8"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("800"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("800"),
        )

        snapshot = _bi_daily_sales_snapshot()

        self.assertEqual(snapshot["comparison_label"], "Arriba")
        self.assertEqual(snapshot["comparison_tone"], "success")
        self.assertIn(prev_day.isoformat(), snapshot["comparison_detail"])

    def test_bi_daily_sales_snapshot_keeps_last_closed_cut_when_current_day_has_partial_sales(self):
        sucursal = self._create_sucursal("CUT-LAST", "Sucursal Corte Cerrado")
        closed_cut_day = date(2026, 4, 21)
        partial_day = date(2026, 4, 22)

        CorteOficialDiario.objects.create(
            corte_date=closed_cut_day,
            total_amount=Decimal("92040.99"),
            total_tickets=310,
            avg_ticket=Decimal("296.91"),
            contado_amount=Decimal("92040.99"),
            credito_amount=Decimal("0"),
            discounts_amount=Decimal("0"),
            new_customers=3,
            evidence_path="/tmp/corte_oficial_2026-04-21.png",
        )
        FactVentaDiaria.objects.create(
            fecha=closed_cut_day,
            sucursal=sucursal,
            producto_clave="CUT-CLOSED",
            producto_nombre="Pastel Corte Cerrado",
            cantidad=Decimal("120"),
            tickets=310,
            venta_bruta=Decimal("92040.99"),
            descuento=Decimal("0"),
            venta_total=Decimal("92040.99"),
            venta_neta=Decimal("92040.99"),
            costo_estimado=Decimal("0"),
            margen=Decimal("0"),
            source_kind=FactVentaDiaria.SOURCE_AUTHORITATIVE,
        )
        FactVentaDiaria.objects.create(
            fecha=partial_day,
            sucursal=sucursal,
            producto_clave="CUT-PARTIAL",
            producto_nombre="Pastel Parcial Dia Actual",
            cantidad=Decimal("60"),
            tickets=150,
            venta_bruta=Decimal("47287.00"),
            descuento=Decimal("0"),
            venta_total=Decimal("47287.00"),
            venta_neta=Decimal("47287.00"),
            costo_estimado=Decimal("0"),
            margen=Decimal("0"),
            source_kind=FactVentaDiaria.SOURCE_AUTHORITATIVE,
        )

        with patch("reportes.dashboard_sales_dataset.timezone.localdate", return_value=partial_day), patch(
            "reportes.views.timezone.localdate", return_value=partial_day
        ):
            snapshot = _bi_daily_sales_snapshot()

        self.assertEqual(snapshot["date"], closed_cut_day)
        self.assertEqual(snapshot["date_label"], "2026-04-21")
        self.assertEqual(snapshot["total_amount"], Decimal("92040.99"))
        self.assertEqual(snapshot["raw_total_amount"], Decimal("92040.99"))

    def test_bi_daily_sales_snapshot_clamps_to_operational_cutoff_without_official_cut_row(self):
        sucursal = self._create_sucursal("CUT-OPS", "Sucursal Corte Operativo")
        closed_cut_day = date(2026, 4, 21)
        partial_day = date(2026, 4, 22)

        FactVentaDiaria.objects.create(
            fecha=closed_cut_day,
            sucursal=sucursal,
            producto_clave="OPS-CLOSED",
            producto_nombre="Pastel Dia Cerrado",
            cantidad=Decimal("110"),
            tickets=280,
            venta_bruta=Decimal("90500.00"),
            descuento=Decimal("0"),
            venta_total=Decimal("90500.00"),
            venta_neta=Decimal("90500.00"),
            costo_estimado=Decimal("0"),
            margen=Decimal("0"),
            source_kind=FactVentaDiaria.SOURCE_AUTHORITATIVE,
        )
        FactVentaDiaria.objects.create(
            fecha=partial_day,
            sucursal=sucursal,
            producto_clave="OPS-PARTIAL",
            producto_nombre="Pastel Dia Parcial",
            cantidad=Decimal("55"),
            tickets=145,
            venta_bruta=Decimal("47287.00"),
            descuento=Decimal("0"),
            venta_total=Decimal("47287.00"),
            venta_neta=Decimal("47287.00"),
            costo_estimado=Decimal("0"),
            margen=Decimal("0"),
            source_kind=FactVentaDiaria.SOURCE_AUTHORITATIVE,
        )

        mocked_now = datetime(2026, 4, 22, 16, 40, tzinfo=timezone.get_current_timezone())
        with patch("reportes.dashboard_sales_dataset.timezone.localtime", return_value=mocked_now), patch(
            "reportes.dashboard_sales_dataset.timezone.localdate", return_value=partial_day
        ), patch("reportes.views.timezone.localdate", return_value=partial_day):
            snapshot = _bi_daily_sales_snapshot()

        self.assertEqual(snapshot["date"], closed_cut_day)
        self.assertEqual(snapshot["date_label"], "2026-04-21")
        self.assertEqual(snapshot["total_amount"], Decimal("90500.00"))
        self.assertEqual(snapshot["raw_total_amount"], Decimal("90500.00"))

    def test_financial_view_renders(self):
        sucursal = self._create_sucursal("FIN-01", "Sucursal Finanzas")
        receta = Receta.objects.create(
            nombre="Pastel Finanzas",
            hash_contenido="hash-finanzas-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            familia="Pasteles",
            categoria="Mediano",
        )
        fecha_actual = timezone.localdate() - timedelta(days=1)
        point_branch = PointBranch.objects.create(external_id="FIN-01", name="Sucursal Finanzas", erp_branch=sucursal)
        point_product = PointProduct.objects.create(external_id="PFIN-01", sku="FIN001", name="Pastel Finanzas", active=True)
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            receta=receta,
            sale_date=fecha_actual,
            quantity=Decimal("10"),
            total_amount=Decimal("2500"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        RecetaCostoSemanal.objects.create(
            scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
            identity_key=f"recipe:{receta.id}",
            label=receta.nombre,
            week_start=fecha_actual - timedelta(days=fecha_actual.weekday()),
            week_end=fecha_actual - timedelta(days=fecha_actual.weekday()) + timedelta(days=6),
            receta=receta,
            temporalidad=Receta.TEMPORALIDAD_PERMANENTE,
            familia="Pasteles",
            categoria="Mediano",
            costo_mp=Decimal("150"),
            costo_total=Decimal("150"),
            delta_total=Decimal("10"),
            delta_pct=Decimal("7.14"),
        )

        resp = self.client.get(reverse("reportes:costo_receta"))

        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Finanzas de producción y margen MP")
        self.assertContains(resp, "Trayectoria semanal de costo")
        self.assertContains(resp, "Margen vs volumen")
        self.assertContains(resp, "Brecha ASP vs precio sugerido")
        self.assertContains(resp, "Pendientes reales de costeo")
        self.assertContains(resp, "financeWeeklyCostChart")
        self.assertContains(resp, "financeProfitabilityChart")
        self.assertContains(resp, "financeDeltaChart")
        self.assertContains(resp, "financeFamilyChart")
        self.assertContains(resp, "financePriceGapChart")
        self.assertContains(resp, "financeBucketChart")
        self.assertNotContains(resp, "Centro de mando ERP")
        self.assertNotContains(resp, "Workflow ERP del módulo")
        self.assertNotContains(resp, "Mesa de gobierno ERP")
        self.assertNotContains(resp, "Ruta crítica ERP")

    def test_bi_shows_plan_supply_watchlist(self):
        sucursal = self._create_sucursal("BI-SUP-01", "Sucursal BI Supply")
        unidad = UnidadMedida.objects.create(
            codigo="g",
            nombre="Gramo BI Supply",
            tipo=UnidadMedida.TIPO_MASA,
        )
        insumo = Insumo.objects.create(
            nombre="Chocolate BI Supply",
            codigo_point="BI-SUP-INS-01",
            unidad_base=unidad,
            tipo_item=Insumo.TIPO_MATERIA_PRIMA,
            activo=True,
        )
        receta = Receta.objects.create(
            nombre="Pastel BI Supply",
            hash_contenido="hash-bi-supply",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=insumo,
            insumo_texto=insumo.nombre,
            unidad=unidad,
            unidad_texto="kg",
            cantidad=Decimal("2000"),
            match_status=LineaReceta.STATUS_AUTO,
        )
        plan = PlanProduccion.objects.create(
            nombre="Plan BI Supply",
            fecha_produccion=timezone.localdate(),
        )
        PlanProduccionItem.objects.create(plan=plan, receta=receta, cantidad=Decimal("4"))
        ExistenciaInsumo.objects.create(insumo=insumo, stock_actual=Decimal("1"), punto_reorden=Decimal("2"))
        ExistenciaInsumo.objects.create(insumo=insumo, almacen="ARMADO", stock_actual=Decimal("3"))
        point_cedis = PointBranch.objects.create(external_id="BI-SUP-CEDIS", name="CEDIS")
        canonical_job = PointSyncJob.objects.create(
            job_type=PointSyncJob.JOB_TYPE_INVENTORY,
            status=PointSyncJob.STATUS_SUCCESS,
            parameters={"canonical_insumo_inventory": True, "locations": ["ALMACEN", "CEDIS"]},
        )
        PointInsumoInventorySnapshot.objects.create(
            branch=point_cedis,
            insumo=insumo,
            point_code=insumo.codigo_point,
            point_name=insumo.nombre,
            point_quantity=Decimal("4"),
            point_unit="kg",
            quantity_base=Decimal("4000"),
            captured_at=timezone.now(),
            sync_job=canonical_job,
        )
        VentaHistorica.objects.create(
            receta=receta,
            sucursal=sucursal,
            fecha=timezone.localdate() - timedelta(days=2),
            cantidad=Decimal("25"),
            monto_total=Decimal("500"),
            fuente="BI_SUPPLY_TEST",
        )

        resp = self.client.get(reverse("reportes:bi"))

        self.assertEqual(resp.status_code, 200)
        self.assertIn("supply_watchlist", resp.context)
        self.assertTrue(resp.context["supply_watchlist"])
        self.assertEqual(resp.context["supply_watchlist"]["plan_nombre"], "Plan BI Supply")
        self.assertEqual(resp.context["supply_watchlist"]["rows"][0]["insumo_nombre"], "Chocolate BI Supply")
        self.assertEqual(resp.context["supply_watchlist"]["rows"][0]["stock_actual"], Decimal("4000"))
        self.assertEqual(resp.context["supply_watchlist"]["rows"][0]["shortage"], Decimal("4000"))
        self.assertEqual(resp.context["supply_watchlist"]["rows"][0]["stock_actual_display"], Decimal("4"))
        self.assertEqual(resp.context["supply_watchlist"]["rows"][0]["shortage_display"], Decimal("4"))
        self.assertEqual(resp.context["supply_watchlist"]["rows"][0]["display_unit"], "kg")
        self.assertContains(resp, "kg")
        self.assertNotContains(resp, "4,000.00")

    def test_bi_production_summary_uses_bulk_recipe_cost_map(self):
        target_date = timezone.localdate()
        receta = Receta.objects.create(
            nombre="Pastel BI Produccion",
            hash_contenido="hash-bi-production-summary-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        plan = PlanProduccion.objects.create(nombre="Plan BI Produccion", fecha_produccion=target_date)
        PlanProduccionItem.objects.create(plan=plan, receta=receta, cantidad=Decimal("3"))
        VentaHistorica.objects.create(
            receta=receta,
            fecha=target_date,
            cantidad=Decimal("2"),
            monto_total=Decimal("200"),
            fuente="BI_PRODUCTION_SUMMARY_TEST",
        )

        with patch("reportes.views.get_total_cost_map", return_value={receta.id: Decimal("12.50")}) as cost_map:
            summary = _bi_production_summary(target_date, target_date)

        cost_map.assert_called_once_with({receta.id})
        self.assertEqual(summary["total_units"], Decimal("3"))
        self.assertEqual(summary["total_cost"], Decimal("37.50"))
        self.assertEqual(summary["top_products"][0]["cost"], Decimal("37.50"))

    def _create_sucursal(self, codigo: str, nombre: str):
        from core.models import Sucursal

        return Sucursal.objects.create(codigo=codigo, nombre=nombre, activa=True)


class PointSalesMonthTotalTests(TestCase):
    def setUp(self):
        from core.models import Sucursal

        self.sucursal = Sucursal.objects.create(codigo="MONTH-TOTAL", nombre="Sucursal Month Total", activa=True)
        self.point_branch = PointBranch.objects.create(external_id="MONTH-TOTAL", name=self.sucursal.nombre, erp_branch=self.sucursal)
        self.point_product = PointProduct.objects.create(
            external_id="PMONTH",
            sku="PMONTH",
            name="Pastel Month",
            category="Pasteles",
            active=True,
        )
        self.receta = Receta.objects.create(
            nombre="Pastel Month",
            codigo_point="PMONTH",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-point-month-total-001",
        )

    def test_month_total_prefers_canonical_sales_range_over_legacy_sources(self):
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product,
            sale_date=date(2026, 3, 16),
            quantity=Decimal("50"),
            total_amount=Decimal("5000"),
            gross_amount=Decimal("5000"),
            net_amount=Decimal("5000"),
            source_endpoint="/Report/VentasCategorias",
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=self.point_branch,
            sale_date=date(2026, 3, 16),
            sucursal_nombre=self.sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("10"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("1850"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("1850"),
        )
        VentaHistorica.objects.create(
            sucursal=self.sucursal,
            fecha=date(2026, 3, 20),
            receta=self.receta,
            cantidad=Decimal("20"),
            monto_total=Decimal("9000"),
            fuente="POINT_BRIDGE_SALES",
        )

        core_resolved = core_point_sales_month_total(2026, 3)
        reportes_resolved = reportes_point_sales_month_total(2026, 3)

        self.assertEqual(core_resolved["value"], Decimal("1850"))
        self.assertEqual(core_resolved["source_label"], "Point directo")
        self.assertEqual(reportes_resolved["value"], Decimal("1850"))
        self.assertEqual(reportes_resolved["source_label"], "Point directo")

    def test_month_total_preserves_bridge_fallback_when_canonical_has_no_data(self):
        VentaHistorica.objects.create(
            sucursal=self.sucursal,
            fecha=date(2026, 2, 10),
            receta=self.receta,
            cantidad=Decimal("12"),
            monto_total=Decimal("2400"),
            fuente="POINT_BRIDGE_SALES",
        )

        core_resolved = core_point_sales_month_total(2026, 2)
        reportes_resolved = reportes_point_sales_month_total(2026, 2)

        self.assertEqual(core_resolved["value"], Decimal("2400"))
        self.assertEqual(core_resolved["source_label"], "Point conciliado")
        self.assertEqual(reportes_resolved["value"], Decimal("2400"))
        self.assertEqual(reportes_resolved["source_label"], "Point conciliado")


class ReportesBIUtilsTests(TestCase):
    def setUp(self):
        from core.models import Sucursal

        _clear_reportes_runtime_caches()
        # La vista BI exige login + RBAC (can_view_reportes) desde el
        # endurecimiento de accesos; los exports comparten esa vista.
        self.user = User.objects.create_user(username="lectura_bi_utils", password="pass123")
        group, _ = Group.objects.get_or_create(name="LECTURA")
        self.user.groups.add(group)
        self.client.login(username="lectura_bi_utils", password="pass123")
        self.sucursal = Sucursal.objects.create(codigo="BIUTILS", nombre="Sucursal BI Utils", activa=True)
        self.point_branch = PointBranch.objects.create(external_id="BIUTILS", name=self.sucursal.nombre, erp_branch=self.sucursal)
        self.point_product = PointProduct.objects.create(
            external_id="PBIUTILS",
            sku="BIUTILS01",
            name="Pastel BI Utils",
            category="Pasteles",
            active=True,
        )
        self.receta = Receta.objects.create(
            nombre="Pastel BI Utils",
            codigo_point="BIUTILS01",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-bi-utils-001",
        )

    def test_compute_bi_snapshot_agrupa_stock_y_usa_umbral_de_almacen_1(self):
        unidad = UnidadMedida.objects.create(
            codigo="kg-bi-utils-multi",
            nombre="Kilogramo BI Utils Multi",
            tipo=UnidadMedida.TIPO_MASA,
        )
        insumo = Insumo.objects.create(
            nombre="Insumo BI Utils Multi",
            unidad_base=unidad,
            activo=True,
        )
        ExistenciaInsumo.objects.create(
            insumo=insumo,
            almacen="ALMACEN_1",
            stock_actual=Decimal("2"),
            punto_reorden=Decimal("5"),
        )
        ExistenciaInsumo.objects.create(
            insumo=insumo,
            almacen="ARMADO",
            stock_actual=Decimal("4"),
            punto_reorden=Decimal("100"),
        )

        snapshot = compute_bi_snapshot(period_days=7, months_window=3)

        self.assertEqual(snapshot["kpis"]["alertas_stock"], 0)
        self.assertEqual(snapshot["kpis"]["criticos_stock"], 0)
        self.assertEqual(snapshot["kpis"]["bajo_reorden_stock"], 0)
        self.assertEqual(snapshot["kpis"]["inventario_point_no_disponible"], 1)

    def test_compute_bi_snapshot_uses_prefer_complete_canonical_sales_range(self):
        today = timezone.localdate()
        first_day = today - timedelta(days=2)
        second_day = today - timedelta(days=1)

        VentaAutoritativaPoint.objects.create(
            branch=self.sucursal,
            product=self.receta,
            sale_date=first_day,
            product_code="BIUTILS01",
            point_name=self.receta.nombre,
            category="Pasteles",
            quantity=Decimal("4"),
            total_amount=Decimal("400"),
            net_amount=Decimal("400"),
            tax_amount=Decimal("0"),
            gross_amount=Decimal("400"),
            discount_amount=Decimal("0"),
            source_file="test",
            source_sheet="Sheet1",
            raw_payload={},
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=self.point_branch,
            sale_date=first_day,
            sucursal_nombre=self.sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("5"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("500"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("500"),
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=self.point_branch,
            sale_date=second_day,
            sucursal_nombre=self.sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("7"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("700"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("700"),
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product,
            receta=self.receta,
            sale_date=second_day,
            quantity=Decimal("20"),
            total_amount=Decimal("2000"),
            gross_amount=Decimal("2000"),
            net_amount=Decimal("2000"),
            source_endpoint="/Report/VentasCategorias",
        )

        snapshot = compute_bi_snapshot(period_days=7, months_window=3)

        self.assertEqual(snapshot["kpis"]["ventas_total"], Decimal("1200"))
        self.assertEqual(snapshot["kpis"]["pedidos_venta"], 2)
        current_period = f"{second_day.year:04d}-{second_day.month:02d}"
        current_month_row = next(row for row in snapshot["series_mensual"] if row["periodo"] == current_period)
        self.assertEqual(current_month_row["ventas"], Decimal("1200"))
        self.assertFalse(snapshot["kpis"]["official_sales_series_ready"])

    def _create_sucursal(self, codigo: str, nombre: str):
        from core.models import Sucursal

        return Sucursal.objects.create(codigo=codigo, nombre=nombre, activa=True)


    def test_bi_exports(self):
        resp_csv = self.client.get(reverse("reportes:bi"), {"export": "csv"})
        self.assertEqual(resp_csv.status_code, 200)
        self.assertIn("text/csv", resp_csv["Content-Type"])

        resp_xlsx = self.client.get(reverse("reportes:bi"), {"export": "xlsx"})
        self.assertEqual(resp_xlsx.status_code, 200)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", resp_xlsx["Content-Type"])

    def test_requires_login(self):
        self.client.logout()
        resp = self.client.get(reverse("reportes:bi"))
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/login/", resp.url)

    def test_yoy_panel_uses_cached_official_period_for_partial_previous_year(self):
        PointMonthlySalesOfficial.objects.create(
            month_start=timezone.datetime(2025, 10, 1).date(),
            month_end=timezone.datetime(2025, 10, 31).date(),
            total_quantity=Decimal("26971"),
            gross_amount=Decimal("3448997.00"),
            discount_amount=Decimal("3747.00"),
            total_amount=Decimal("3445250.00"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("3406462.00"),
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=timezone.datetime(2025, 11, 1).date(),
            month_end=timezone.datetime(2025, 11, 30).date(),
            total_quantity=Decimal("23292"),
            gross_amount=Decimal("3300000.00"),
            discount_amount=Decimal("53007.52"),
            total_amount=Decimal("3246992.48"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("3200000.00"),
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=timezone.datetime(2025, 12, 1).date(),
            month_end=timezone.datetime(2025, 12, 31).date(),
            total_quantity=Decimal("30202"),
            gross_amount=Decimal("4900000.00"),
            discount_amount=Decimal("75531.09"),
            total_amount=Decimal("4824468.91"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("4700000.00"),
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=timezone.datetime(2026, 1, 1).date(),
            month_end=timezone.datetime(2026, 1, 31).date(),
            total_quantity=Decimal("25644"),
            gross_amount=Decimal("3550000.00"),
            discount_amount=Decimal("30623.99"),
            total_amount=Decimal("3519376.01"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("3519376.01"),
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=timezone.datetime(2026, 2, 1).date(),
            month_end=timezone.datetime(2026, 2, 28).date(),
            total_quantity=Decimal("25786"),
            gross_amount=Decimal("3320000.00"),
            discount_amount=Decimal("31321.81"),
            total_amount=Decimal("3288678.19"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("3288678.19"),
        )
        PointMonthlySalesOfficial.objects.create(
            month_start=timezone.datetime(2025, 3, 1).date(),
            month_end=timezone.datetime(2025, 3, 31).date(),
            total_quantity=Decimal("30785"),
            gross_amount=Decimal("3900000.00"),
            discount_amount=Decimal("43259.29"),
            total_amount=Decimal("3856740.71"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("3810000.00"),
            raw_payload={
                "partial_ranges": {
                    "2025-03-01_2025-03-21": {
                        "period_start": "2025-03-01",
                        "period_end": "2025-03-21",
                        "total_quantity": "22000",
                        "total_amount": "3333333.33",
                        "gross_amount": "3400000.00",
                        "discount_amount": "50000.00",
                        "tax_amount": "0",
                        "net_amount": "3300000.00",
                    }
                }
            },
        )
        sucursal = self._create_sucursal("BI-PARTIAL", "Sucursal Parcial")
        point_branch = PointBranch.objects.create(external_id="BI-PARTIAL", name="Sucursal Parcial", erp_branch=sucursal)
        point_product = PointProduct.objects.create(external_id="PBI-PARTIAL", sku="BI-PARTIAL", name="Producto parcial", active=True)
        fecha_actual = timezone.datetime(2026, 3, 21).date()
        for idx in range(21):
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product,
                sale_date=fecha_actual.replace(day=1) + timedelta(days=idx),
                quantity=Decimal("10"),
                total_amount=Decimal("126829.35"),
                source_endpoint="/Report/PrintReportes?idreporte=3",
            )

        panel = build_monthly_yoy_panel(latest_date=fecha_actual, months=6)
        march_row = panel["rows"][-1]
        self.assertEqual(march_row["month_label"], "2026-03")
        self.assertEqual(march_row["prev_amount"], Decimal("3333333.33"))

    def test_yoy_panel_uses_canonical_range_when_monthly_cache_missing(self):
        fecha_actual = date(2026, 3, 31)
        sucursal = self._create_sucursal("BI-YOY-CAN", "Sucursal YOY Can")
        point_branch = PointBranch.objects.create(external_id="BI-YOY-CAN", name="Sucursal YOY Can", erp_branch=sucursal)
        point_product = PointProduct.objects.create(external_id="PBI-YOY-CAN", sku="BIYOYCAN", name="Producto YOY Can", active=True)
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            sale_date=date(2026, 3, 16),
            quantity=Decimal("50"),
            total_amount=Decimal("5000"),
            source_endpoint="/Report/VentasCategorias",
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=point_branch,
            sale_date=date(2026, 3, 16),
            sucursal_nombre=sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("10"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("1850"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("1850"),
        )

        panel = build_monthly_yoy_panel(latest_date=fecha_actual, months=1)

        march_row = panel["rows"][0]
        self.assertEqual(march_row["month_label"], "2026-03")
        self.assertEqual(march_row["amount"], Decimal("1850.00"))
        self.assertEqual(march_row["quantity"], Decimal("10.00"))

    def test_month_total_ignores_duplicated_fact_table_when_point_is_available(self):
        FactVentaDiaria.objects.create(
            fecha=date(2026, 3, 16),
            sucursal=self.sucursal,
            receta=self.receta,
            producto_clave="BIUTILS01",
            producto_nombre="Pastel BI Utils",
            cantidad=Decimal("20"),
            tickets=10,
            venta_bruta=Decimal("3700"),
            descuento=Decimal("0"),
            venta_total=Decimal("3700"),
            venta_neta=Decimal("3700"),
            costo_estimado=Decimal("0"),
            margen=Decimal("0"),
            source_kind=FactVentaDiaria.SOURCE_AUTHORITATIVE,
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=self.point_branch,
            sale_date=date(2026, 3, 16),
            sucursal_nombre=self.sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("10"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("1850"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("1850"),
        )

        resolved = reportes_point_sales_month_total(2026, 3)

        self.assertEqual(resolved["value"], Decimal("1850"))
        self.assertEqual(resolved["source_label"], "Point directo")

    def test_yoy_panel_ignores_duplicated_fact_table_when_point_is_available(self):
        fecha_actual = date(2026, 3, 31)
        FactVentaDiaria.objects.create(
            fecha=date(2026, 3, 16),
            sucursal=self.sucursal,
            receta=self.receta,
            producto_clave="BIUTILS01",
            producto_nombre="Pastel BI Utils",
            cantidad=Decimal("20"),
            tickets=10,
            venta_bruta=Decimal("3700"),
            descuento=Decimal("0"),
            venta_total=Decimal("3700"),
            venta_neta=Decimal("3700"),
            costo_estimado=Decimal("0"),
            margen=Decimal("0"),
            source_kind=FactVentaDiaria.SOURCE_AUTHORITATIVE,
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=self.point_branch,
            sale_date=date(2026, 3, 16),
            sucursal_nombre=self.sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("10"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("1850"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("1850"),
        )

        panel = build_monthly_yoy_panel(latest_date=fecha_actual, months=1)

        self.assertEqual(panel["rows"][0]["amount"], Decimal("1850.00"))
        self.assertEqual(panel["rows"][0]["quantity"], Decimal("10.00"))

    @patch(
        "reportes.executive_panels._best_partial_cache_payload",
        return_value={
            "period_start": "2026-03-01",
            "period_end": "2026-03-23",
            "total_amount": "1800",
            "total_quantity": "9",
        },
    )
    def test_yoy_panel_ignores_partial_cache_when_month_is_closed(self, _partial_cache):
        fecha_actual = date(2026, 3, 31)
        PointSalesDailyCategoryFact.objects.create(
            branch=self.point_branch,
            sale_date=date(2026, 3, 16),
            sucursal_nombre=self.sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("10"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("1850"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("1850"),
        )

        panel = build_monthly_yoy_panel(latest_date=fecha_actual, months=1)

        self.assertEqual(panel["rows"][0]["amount"], Decimal("1850.00"))
        self.assertEqual(panel["rows"][0]["quantity"], Decimal("10.00"))

    def test_dashboard_monthly_rows_ignore_duplicated_fact_table_when_point_is_available(self):
        fecha_actual = date(2026, 3, 31)
        FactVentaDiaria.objects.create(
            fecha=date(2026, 3, 16),
            sucursal=self.sucursal,
            receta=self.receta,
            producto_clave="BIUTILS01",
            producto_nombre="Pastel BI Utils",
            cantidad=Decimal("20"),
            tickets=10,
            venta_bruta=Decimal("3700"),
            descuento=Decimal("0"),
            venta_total=Decimal("3700"),
            venta_neta=Decimal("3700"),
            costo_estimado=Decimal("0"),
            margen=Decimal("0"),
            source_kind=FactVentaDiaria.SOURCE_AUTHORITATIVE,
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=self.point_branch,
            sale_date=date(2026, 3, 16),
            sucursal_nombre=self.sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("10"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("1850"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("1850"),
        )

        with patch("reportes.dashboard_sales_dataset.timezone.localdate", return_value=fecha_actual):
            dataset = get_dashboard_sales_dataset(today=fecha_actual, months=1)

        self.assertEqual(dataset["monthly_sales_rows"][0]["value"], Decimal("1850"))
        self.assertEqual(dataset["daily_sales_snapshot"]["month_amount"], Decimal("1850"))

    def test_monthly_point_totals_ignore_duplicated_autoritative_table_when_stage_sources_match(self):
        fecha_venta = date(2026, 3, 16)
        VentaAutoritativaPoint.objects.create(
            branch=self.sucursal,
            product=self.receta,
            sale_date=fecha_venta,
            product_code="BIUTILS01",
            point_name=self.receta.nombre,
            category="Pasteles",
            quantity=Decimal("20"),
            gross_amount=Decimal("3700"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("3700"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("3600"),
            source_file="dup-a",
            source_sheet="Sheet1",
            raw_payload={},
        )
        VentaAutoritativaPoint.objects.create(
            branch=self.sucursal,
            product=self.receta,
            sale_date=fecha_venta,
            product_code="BIUTILS01-DUP",
            point_name=self.receta.nombre,
            category="Pasteles",
            quantity=Decimal("20"),
            gross_amount=Decimal("3700"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("3700"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("3600"),
            source_file="dup-b",
            source_sheet="Sheet1",
            raw_payload={},
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=self.point_branch,
            sale_date=fecha_venta,
            sucursal_nombre=self.sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("10"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("1850"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("1800"),
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=self.point_product,
            receta=self.receta,
            sale_date=fecha_venta,
            quantity=Decimal("10"),
            gross_amount=Decimal("1850"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("1850"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("1800"),
            source_endpoint="/Report/VentasCategorias",
        )

        resolved = reportes_point_sales_month_total(2026, 3)
        panel = build_monthly_yoy_panel(latest_date=date(2026, 3, 31), months=1)

        self.assertEqual(resolved["value"], Decimal("1850"))
        self.assertEqual(panel["rows"][0]["amount"], Decimal("1850.00"))
        self.assertEqual(panel["rows"][0]["quantity"], Decimal("10.00"))

    def test_central_flow_panel_uses_canonical_range_when_monthly_cache_missing(self):
        fecha_actual = date(2026, 3, 31)
        sucursal = self._create_sucursal("BI-FLOW-CAN", "Sucursal Flow Can")
        point_branch = PointBranch.objects.create(external_id="BI-FLOW-CAN", name="Sucursal Flow Can", erp_branch=sucursal)
        point_product = PointProduct.objects.create(external_id="PBI-FLOW-CAN", sku="BIFLOWCAN", name="Producto Flow Can", active=True)
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            sale_date=date(2026, 3, 18),
            quantity=Decimal("50"),
            total_amount=Decimal("5000"),
            source_endpoint="/Report/VentasCategorias",
        )
        PointSalesDailyCategoryFact.objects.create(
            branch=point_branch,
            sale_date=date(2026, 3, 18),
            sucursal_nombre=sucursal.nombre,
            categoria="Pasteles",
            total_cantidad=Decimal("12"),
            total_descuento=Decimal("0"),
            total_venta=Decimal("2400"),
            total_impuestos=Decimal("0"),
            total_venta_neta=Decimal("2400"),
        )

        panel = build_central_flow_panel(latest_date=fecha_actual, months=1)

        month_row = panel["rows"][0]
        self.assertEqual(month_row["month_label"], "2026-03")
        self.assertEqual(month_row["sold_units"], Decimal("12.00"))


class ReportesCanonicosTests(TestCase):
    def setUp(self):
        _clear_reportes_runtime_caches()
        self.user = User.objects.create_user(username="lectura_reportes_cat", password="pass123")
        group, _ = Group.objects.get_or_create(name="LECTURA")
        self.user.groups.add(group)
        self.client.login(username="lectura_reportes_cat", password="pass123")

    def test_consumo_agrupa_movimientos_de_variantes_en_canonico(self):
        unidad = UnidadMedida.objects.create(codigo="kg-rpt", nombre="Kg Reporte", tipo=UnidadMedida.TIPO_MASA)
        canonical = Insumo.objects.create(
            nombre="Harina Canonica Reporte",
            unidad_base=unidad,
            activo=True,
            codigo_point="RPT-CAN-001",
        )
        variant = Insumo.objects.create(
            nombre="HARINA CANONICA REPORTE",
            unidad_base=unidad,
            activo=True,
        )
        MovimientoInventario.objects.create(
            tipo=MovimientoInventario.TIPO_SALIDA,
            insumo=canonical,
            cantidad=Decimal("2"),
            referencia="RPT-1",
        )
        MovimientoInventario.objects.create(
            tipo=MovimientoInventario.TIPO_SALIDA,
            insumo=variant,
            cantidad=Decimal("3"),
            referencia="RPT-2",
        )

        response = self.client.get(reverse("reportes:consumo"))
        self.assertEqual(response.status_code, 200)
        rows = [row for row in response.context["rows"] if row["insumo__nombre"] == canonical.nombre]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["cantidad_total"], Decimal("5"))
        self.assertEqual(response.context["total_insumos"], 1)
        self.assertContains(response, "Maestro listo")
        self.assertContains(response, "Maestro ERP")
        self.assertContains(response, "Cadena de control")
        self.assertContains(response, "Cierre por etapa")
        self.assertContains(response, "Responsable")
        self.assertContains(response, "Cierre")
        self.assertContains(response, "Depende de")
        self.assertContains(response, "Dependencia")
        self.assertIn("maturity_summary", response.context)
        self.assertIn("handoff_map", response.context)
        self.assertIn("release_gate_rows", response.context)
        self.assertIn("release_gate_completion", response.context)
        self.assertIn("dependency_status", response.context["enterprise_chain"][0])
        self.assertIn("owner", response.context["document_stage_rows"][0])
        self.assertIn("completion", response.context["document_stage_rows"][0])
        self.assertIn("erp_governance_rows", response.context)
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertIn("erp_command_center", response.context)

    def test_faltantes_agrupa_existencias_de_variantes_en_canonico(self):
        unidad = UnidadMedida.objects.create(codigo="pz-rpt", nombre="Pza Reporte", tipo=UnidadMedida.TIPO_PIEZA)
        canonical = Insumo.objects.create(
            nombre="Caja Canonica Reporte",
            unidad_base=unidad,
            activo=True,
            codigo_point="RPT-EX-001",
        )
        variant = Insumo.objects.create(
            nombre="CAJA CANONICA REPORTE",
            unidad_base=unidad,
            activo=True,
        )
        ExistenciaInsumo.objects.create(insumo=canonical, stock_actual=Decimal("4"), punto_reorden=Decimal("10"))
        ExistenciaInsumo.objects.create(insumo=variant, stock_actual=Decimal("3"), punto_reorden=Decimal("2"))
        point_almacen = PointBranch.objects.create(external_id="RPT-EX-ALMACEN", name="Almacen")
        canonical_job = PointSyncJob.objects.create(
            job_type=PointSyncJob.JOB_TYPE_INVENTORY,
            status=PointSyncJob.STATUS_SUCCESS,
            parameters={"canonical_insumo_inventory": True, "locations": ["ALMACEN", "CEDIS"]},
        )
        PointInsumoInventorySnapshot.objects.create(
            branch=point_almacen,
            insumo=canonical,
            point_code=canonical.codigo_point,
            point_name=canonical.nombre,
            point_quantity=Decimal("7"),
            point_unit="pza",
            quantity_base=Decimal("7"),
            captured_at=timezone.now(),
            sync_job=canonical_job,
        )

        response = self.client.get(reverse("reportes:faltantes"), {"nivel": "all"})
        self.assertEqual(response.status_code, 200)
        rows = [row for row in response.context["rows"] if row.insumo.nombre == canonical.nombre]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].stock_actual, Decimal("7"))
        self.assertContains(response, "Maestro bloqueado")
        self.assertContains(response, "Maestro ERP")
        self.assertContains(response, "Cadena de control")
        self.assertContains(response, "Cierre por etapa")
        self.assertContains(response, "Responsable")
        self.assertContains(response, "Cierre")
        self.assertContains(response, "Depende de")
        self.assertContains(response, "Dependencia")
        self.assertIn("maturity_summary", response.context)
        self.assertIn("handoff_map", response.context)
        self.assertIn("release_gate_rows", response.context)
        self.assertIn("release_gate_completion", response.context)
        self.assertIn("dependency_status", response.context["enterprise_chain"][0])
        self.assertIn("owner", response.context["document_stage_rows"][0])
        self.assertIn("completion", response.context["document_stage_rows"][0])
        self.assertIn("erp_governance_rows", response.context)
        self.assertIn("critical_path_rows", response.context)
        self.assertIn("executive_radar_rows", response.context)
        self.assertIn("erp_command_center", response.context)

    def test_costo_receta_usa_costo_canonico_para_variante_historica(self):
        unidad = UnidadMedida.objects.create(codigo="kg-rpt-cost", nombre="Kg Reporte Costo", tipo=UnidadMedida.TIPO_MASA)
        canonical = Insumo.objects.create(
            nombre="Crema Canonica Reporte",
            unidad_base=unidad,
            activo=True,
            codigo_point="RPT-COST-001",
        )
        variant = Insumo.objects.create(
            nombre="CREMA CANONICA REPORTE",
            unidad_base=unidad,
            activo=True,
        )
        CostoInsumo.objects.create(insumo=canonical, costo_unitario=Decimal("42.50"))
        receta = Receta.objects.create(
            nombre="Receta Canonica Reporte",
            hash_contenido="hash-rpt-cost-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        LineaReceta.objects.create(
            receta=receta,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=variant,
            insumo_texto=variant.nombre,
            cantidad=Decimal("2.000"),
            unidad_texto="kg",
            match_status=LineaReceta.STATUS_AUTO,
        )

        response = self.client.get(reverse("reportes:costo_receta"))
        self.assertEqual(response.status_code, 200)
        rows = [row for row in response.context["recipe_rows"] if row["receta"].id == receta.id]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["costo_total"], Decimal("85.00"))
        self.assertEqual(rows[0]["lineas_costeadas"], 1)
        self.assertContains(response, "Finanzas de producción y margen MP")
        self.assertContains(response, "precio sugerido")
        self.assertContains(response, "Trayectoria semanal de costo")
        self.assertContains(response, "Margen vs volumen")
        self.assertContains(response, "Pendientes reales de costeo")
        self.assertIn("current_week_rows", response.context)
        self.assertIn("profitability_panel", response.context)
        self.assertIn("recipe_rows", response.context)

    def test_costo_receta_filters_by_family_and_query(self):
        week_start = timezone.localdate() - timedelta(days=timezone.localdate().weekday())
        pastel = Receta.objects.create(
            nombre="Pastel Filtro Costo",
            hash_contenido="hash-cost-filter-1",
            familia="Pasteles",
            categoria="Mediano",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        pay = Receta.objects.create(
            nombre="Pay Filtro Costo",
            hash_contenido="hash-cost-filter-2",
            familia="Pays",
            categoria="General",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        sucursal = self._create_sucursal("RPT-FILTER", "Sucursal Reporte Filter")
        point_branch = PointBranch.objects.create(external_id="RPT-FILTER", name="Sucursal Reporte Filter", erp_branch=sucursal)
        for receta, delta in ((pastel, Decimal("12")), (pay, Decimal("5"))):
            RecetaCostoSemanal.objects.create(
                scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
                identity_key=f"recipe:{receta.id}",
                label=receta.nombre,
                week_start=week_start,
                week_end=week_start + timedelta(days=6),
                receta=receta,
                temporalidad=Receta.TEMPORALIDAD_PERMANENTE,
                familia=receta.familia,
                categoria=receta.categoria,
                costo_mp=Decimal("100"),
                costo_total=Decimal("100"),
                delta_total=delta,
                delta_pct=Decimal("10"),
            )
            point_product = PointProduct.objects.create(
                external_id=f"RPT-FILTER-{receta.id}",
                sku=f"RPTFILTER{receta.id:02d}",
                name=receta.nombre,
                active=True,
            )
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product,
                sale_date=timezone.localdate() - timedelta(days=1),
                receta=receta,
                quantity=Decimal("10"),
                total_amount=Decimal("2500"),
                source_endpoint="/Report/PrintReportes?idreporte=3",
            )

        response = self.client.get(
            reverse("reportes:costo_receta"),
            {"familia": "Pasteles", "categoria": "Mediano", "q": "Pastel"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_familia"], "Pasteles")
        self.assertEqual(response.context["selected_categoria"], "Mediano")
        self.assertEqual(response.context["selected_q"], "Pastel")
        self.assertEqual(len(response.context["current_week_rows"]), 1)
        self.assertContains(response, 'option value="Pasteles" selected')
        self.assertContains(response, 'option value="Mediano" selected')

    def test_costo_receta_filters_by_bucket(self):
        week_start = timezone.localdate() - timedelta(days=timezone.localdate().weekday())
        receta_baja = Receta.objects.create(
            nombre="Pastel Promo",
            hash_contenido="hash-cost-bucket-1",
            familia="Pasteles",
            categoria="Mediano",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        receta_alta = Receta.objects.create(
            nombre="Pastel Defender",
            hash_contenido="hash-cost-bucket-2",
            familia="Pasteles",
            categoria="Grande",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        sucursal = self._create_sucursal("RPT-BUCKET", "Sucursal Reporte Bucket")
        point_branch = PointBranch.objects.create(external_id="RPT-BUCKET", name="Sucursal Reporte Bucket", erp_branch=sucursal)
        for receta in (receta_baja, receta_alta):
            RecetaCostoSemanal.objects.create(
                scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
                identity_key=f"recipe:{receta.id}",
                label=receta.nombre,
                week_start=week_start,
                week_end=week_start + timedelta(days=6),
                receta=receta,
                temporalidad=Receta.TEMPORALIDAD_PERMANENTE,
                familia=receta.familia,
                categoria=receta.categoria,
                costo_mp=Decimal("100"),
                costo_total=Decimal("100"),
                delta_total=Decimal("5"),
                delta_pct=Decimal("5"),
            )
        for receta, qty, amount in (
            (receta_baja, Decimal("4"), Decimal("1200")),
            (receta_alta, Decimal("20"), Decimal("6000")),
        ):
            point_product = PointProduct.objects.create(
                external_id=f"RPT-BUCKET-{receta.id}",
                sku=f"RPTBUCKET{receta.id:02d}",
                name=receta.nombre,
                active=True,
            )
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product,
                sale_date=timezone.localdate() - timedelta(days=1),
                receta=receta,
                quantity=qty,
                total_amount=amount,
                source_endpoint="/Report/PrintReportes?idreporte=3",
            )

        response = self.client.get(reverse("reportes:costo_receta"), {"bucket": "Promocionar"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_bucket"], "Promocionar")
        self.assertTrue(response.context["profitability_panel"]["rows"])
        self.assertTrue(all(row["bucket"] == "Promocionar" for row in response.context["profitability_panel"]["rows"]))
        self.assertContains(response, 'option value="Promocionar" selected')

    def test_costo_receta_usa_margen_real_contra_venta_y_semaforo_costo(self):
        week_start = timezone.localdate() - timedelta(days=timezone.localdate().weekday())
        receta = Receta.objects.create(
            nombre="Pastel Semaforo Financiero",
            hash_contenido="hash-cost-signal-1",
            familia="Pasteles",
            categoria="Mediano",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        sucursal = self._create_sucursal("RPT-SIGNAL", "Sucursal Reporte Signal")
        point_branch = PointBranch.objects.create(external_id="RPT-SIGNAL", name="Sucursal Reporte Signal", erp_branch=sucursal)
        RecetaCostoSemanal.objects.create(
            scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
            identity_key=f"recipe:{receta.id}",
            label=receta.nombre,
            week_start=week_start,
            week_end=week_start + timedelta(days=6),
            receta=receta,
            temporalidad=Receta.TEMPORALIDAD_PERMANENTE,
            familia=receta.familia,
            categoria=receta.categoria,
            costo_mp=Decimal("100"),
            costo_total=Decimal("100"),
            delta_total=Decimal("5"),
            delta_pct=Decimal("5"),
        )
        point_product = PointProduct.objects.create(
            external_id=f"RPT-SIGNAL-{receta.id}",
            sku=f"RPTSIGNAL{receta.id:02d}",
            name=receta.nombre,
            active=True,
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            sale_date=timezone.localdate() - timedelta(days=1),
            receta=receta,
            quantity=Decimal("10"),
            total_amount=Decimal("2500"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )

        response = self.client.get(reverse("reportes:costo_receta"), {"margen": "35"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["margen_pct"], Decimal("35"))
        self.assertEqual(response.context["avg_cost_pct"], Decimal("40.00"))
        self.assertEqual(response.context["avg_cost_signal"]["label"], "Verde")
        self.assertEqual(response.context["profitability_panel"]["cost_signal_counts"]["success"], 1)
        self.assertTrue(response.context["price_gap_rows"])
        self.assertEqual(response.context["price_gap_rows"][0]["suggested_price"], Decimal("153.85"))
        self.assertEqual(response.context["price_gap_rows"][0]["cost_pct"], Decimal("40.00"))
        self.assertEqual(response.context["profitability_panel"]["rows"][0]["cost_signal_label"], "Verde")
        self.assertTrue(response.context["pricing_action_rows"])
        self.assertEqual(response.context["pricing_action_rows"][0]["suggested_price"], Decimal("153.85"))
        self.assertEqual(response.context["pricing_action_rows"][0]["cost_signal_label"], "Verde")
        self.assertEqual(response.context["pricing_action_rows"][0]["priority_label"], "Monitorear")
        self.assertEqual(response.context["pricing_action_rows"][0]["dg_action_label"], "Defender")
        self.assertIn("dg_action_note", response.context["pricing_action_rows"][0])
        self.assertEqual(response.context["pricing_action_rows"][0]["visible_impact_amount"], Decimal("0.00"))
        self.assertEqual(response.context["pricing_decision_counts"]["defender"], 1)
        self.assertEqual(response.context["pricing_total_visible_impact"], Decimal("0.00"))
        self.assertTrue(response.context["pricing_action_summary_rows"])
        self.assertEqual(response.context["pricing_action_summary_rows"][3]["label"], "Defender")
        self.assertContains(response, "Costo MP / venta")
        self.assertContains(response, "Margen objetivo contra venta")
        self.assertContains(response, "Top 10 para mover primero")
        self.assertContains(response, "Acción DG")
        self.assertContains(response, "Impacto 28d")
        self.assertContains(response, "Glosario ejecutivo")
        self.assertContains(response, "Average Selling Price")

    def test_costo_receta_prefiere_costo_agrupado_si_hay_un_addon_aprobado(self):
        week_start = timezone.localdate() - timedelta(days=timezone.localdate().weekday())
        receta = Receta.objects.create(
            nombre="Pastel Fresa QA",
            codigo_point="PFQA",
            hash_contenido="hash-fin-addon-base-001",
            familia="Pasteles",
            categoria="Mini",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        addon = Receta.objects.create(
            nombre="TOPPING FRESA QA",
            codigo_point="TOPFQA",
            hash_contenido="hash-fin-addon-top-001",
            familia="Pasteles",
            categoria="Mini",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        sucursal = self._create_sucursal("RPT-ADDON", "Sucursal Reporte Addon")
        point_branch = PointBranch.objects.create(external_id="RPT-ADDON", name="Sucursal Reporte Addon", erp_branch=sucursal)
        point_product = PointProduct.objects.create(
            external_id="RPT-ADDON-PROD",
            sku="PFQA",
            name=receta.nombre,
            active=True,
        )
        RecetaCostoSemanal.objects.create(
            scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
            identity_key=f"RECIPE:{receta.id}",
            label=receta.nombre,
            week_start=week_start,
            week_end=week_start + timedelta(days=6),
            receta=receta,
            base_receta=receta,
            temporalidad=Receta.TEMPORALIDAD_PERMANENTE,
            familia=receta.familia,
            categoria=receta.categoria,
            costo_mp=Decimal("50"),
            costo_total=Decimal("50"),
        )
        rule = RecetaAgrupacionAddon.objects.create(
            base_receta=receta,
            addon_receta=addon,
            addon_codigo_point="TOPFQA",
            addon_nombre_point="TOPPING FRESA QA",
            status=RecetaAgrupacionAddon.STATUS_APPROVED,
            activo=True,
        )
        RecetaCostoSemanal.objects.create(
            scope_type=RecetaCostoSemanal.SCOPE_GROUPED_ADDON,
            identity_key=f"GROUPED_ADDON:{rule.id}",
            label=f"{receta.nombre} + TOPPING FRESA QA",
            week_start=week_start,
            week_end=week_start + timedelta(days=6),
            receta=None,
            addon_rule=rule,
            base_receta=receta,
            addon_receta=addon,
            temporalidad=Receta.TEMPORALIDAD_PERMANENTE,
            familia=receta.familia,
            categoria=receta.categoria,
            costo_mp=Decimal("60"),
            costo_total=Decimal("60"),
        )
        PointDailySale.objects.create(
            branch=point_branch,
            product=point_product,
            sale_date=timezone.localdate() - timedelta(days=1),
            receta=receta,
            quantity=Decimal("10"),
            total_amount=Decimal("2000"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )

        response = self.client.get(reverse("reportes:costo_receta"), {"q": "Pastel Fresa QA", "margen": "35"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["profitability_panel"]["rows"][0]["unit_cost"], Decimal("60.00"))
        self.assertEqual(response.context["price_gap_rows"][0]["unit_cost"], Decimal("60.00"))
        self.assertEqual(response.context["price_gap_rows"][0]["suggested_price"], Decimal("92.31"))

    def test_costo_receta_filters_by_coverage(self):
        week_start = timezone.localdate() - timedelta(days=timezone.localdate().weekday())
        unidad = UnidadMedida.objects.create(codigo="kg-rpt-cov", nombre="Kg Rpt Cov", tipo=UnidadMedida.TIPO_MASA)
        insumo_ok = Insumo.objects.create(nombre="Insumo Cobertura OK", unidad_base=unidad, activo=True)
        insumo_partial = Insumo.objects.create(nombre="Insumo Cobertura Parcial", unidad_base=unidad, activo=True)

        receta_completa = Receta.objects.create(
            nombre="Pastel Cobertura Completa",
            hash_contenido="hash-cost-coverage-1",
            familia="Pasteles",
            categoria="Completo",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        receta_parcial = Receta.objects.create(
            nombre="Pastel Cobertura Parcial",
            hash_contenido="hash-cost-coverage-2",
            familia="Pasteles",
            categoria="Parcial",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        receta_bloqueada = Receta.objects.create(
            nombre="Pastel Cobertura Bloqueada",
            hash_contenido="hash-cost-coverage-3",
            familia="Pasteles",
            categoria="Bloqueado",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
        )

        LineaReceta.objects.create(
            receta=receta_completa,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=insumo_ok,
            insumo_texto=insumo_ok.nombre,
            cantidad=Decimal("1"),
            unidad_texto="kg",
            costo_linea_excel=Decimal("80"),
            match_status=LineaReceta.STATUS_AUTO,
        )
        LineaReceta.objects.create(
            receta=receta_parcial,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=insumo_partial,
            insumo_texto=insumo_partial.nombre,
            cantidad=Decimal("1"),
            unidad_texto="kg",
            costo_linea_excel=Decimal("55"),
            match_status=LineaReceta.STATUS_AUTO,
        )
        LineaReceta.objects.create(
            receta=receta_parcial,
            posicion=2,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo=insumo_ok,
            insumo_texto=insumo_ok.nombre,
            cantidad=Decimal("1"),
            unidad_texto="kg",
            match_status=LineaReceta.STATUS_AUTO,
        )
        LineaReceta.objects.create(
            receta=receta_bloqueada,
            posicion=1,
            tipo_linea=LineaReceta.TIPO_NORMAL,
            insumo_texto="Ingrediente sin maestro",
            cantidad=Decimal("1"),
            unidad_texto="kg",
            match_status=LineaReceta.STATUS_NEEDS_REVIEW,
        )

        sucursal = self._create_sucursal("RPT-COVER", "Sucursal Reporte Coverage")
        point_branch = PointBranch.objects.create(external_id="RPT-COVER", name="Sucursal Reporte Coverage", erp_branch=sucursal)
        for receta in (receta_completa, receta_parcial, receta_bloqueada):
            RecetaCostoSemanal.objects.create(
                scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
                identity_key=f"recipe:{receta.id}",
                label=receta.nombre,
                week_start=week_start,
                week_end=week_start + timedelta(days=6),
                receta=receta,
                temporalidad=Receta.TEMPORALIDAD_PERMANENTE,
                familia=receta.familia,
                categoria=receta.categoria,
                costo_mp=Decimal("100"),
                costo_total=Decimal("100"),
                delta_total=Decimal("5"),
                delta_pct=Decimal("5"),
            )
            point_product = PointProduct.objects.create(
                external_id=f"RPT-COVER-{receta.id}",
                sku=f"RPTCOVER{receta.id:02d}",
                name=receta.nombre,
                active=True,
            )
            PointDailySale.objects.create(
                branch=point_branch,
                product=point_product,
                sale_date=timezone.localdate() - timedelta(days=1),
                receta=receta,
                quantity=Decimal("5"),
                total_amount=Decimal("1500"),
                source_endpoint="/Report/PrintReportes?idreporte=3",
            )

        response = self.client.get(reverse("reportes:costo_receta"), {"coverage": "blocked"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_coverage"], "blocked")
        self.assertTrue(response.context["recipe_rows"])
        self.assertIn(
            "Pastel Cobertura Bloqueada",
            [row["receta"].nombre for row in response.context["recipe_rows"]],
        )
        self.assertTrue(response.context["current_week_rows"])
        self.assertIn(
            "Pastel Cobertura Bloqueada",
            [row.label for row in response.context["current_week_rows"]],
        )
        self.assertTrue(response.context["profitability_panel"]["rows"])
        self.assertIn(
            "Pastel Cobertura Bloqueada",
            [row["label"] for row in response.context["profitability_panel"]["rows"]],
        )
        self.assertContains(response, 'option value="blocked" selected')

    def test_costo_receta_supports_base_scope_without_sales_margin(self):
        week_start = timezone.localdate() - timedelta(days=timezone.localdate().weekday())
        receta_base = Receta.objects.create(
            nombre="Batida Base Analitica",
            hash_contenido="hash-cost-base-1",
            familia="Bases",
            categoria="Batidas",
            tipo=Receta.TIPO_PREPARACION,
        )
        RecetaCostoSemanal.objects.create(
            scope_type=RecetaCostoSemanal.SCOPE_RECIPE,
            identity_key=f"recipe:{receta_base.id}",
            label=receta_base.nombre,
            week_start=week_start,
            week_end=week_start + timedelta(days=6),
            receta=receta_base,
            temporalidad=Receta.TEMPORALIDAD_PERMANENTE,
            familia=receta_base.familia,
            categoria=receta_base.categoria,
            costo_mp=Decimal("80"),
            costo_total=Decimal("95"),
            delta_total=Decimal("7"),
            delta_pct=Decimal("8"),
        )

        response = self.client.get(reverse("reportes:costo_receta"), {"scope": "base"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_scope"], "base")
        self.assertFalse(response.context["supports_sales_lens"])
        self.assertEqual(len(response.context["current_week_rows"]), 1)
        self.assertEqual(response.context["current_week_rows"][0].label, "Batida Base Analitica")
        self.assertEqual(response.context["profitability_panel"]["rows"], [])
        self.assertContains(response, 'option value="base" selected')

    def test_cierre_producto_view_renders_existing_closure(self):
        receta = Receta.objects.create(
            nombre="Pastel Cierre Operativo",
            codigo_point="CIE001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-cierre-operativo-001",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2025, 9, 1),
            month_end=date(2025, 9, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
            opening_reference_date=date(2025, 8, 31),
            built_by=self.user,
            notes="Conciliacion teorica mensual.\nSin inventario fisico.",
        )
        line = ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=receta,
            inventario_inicial_teorico=Decimal("15"),
            produccion_mes=Decimal("10"),
            venta_directa_enteros=Decimal("8"),
            venta_derivada_equivalente=Decimal("2"),
            venta_total_equivalente=Decimal("10"),
            merma_total_equivalente=Decimal("1"),
            inventario_final_teorico=Decimal("14"),
            source_snapshot_count=1,
            source_sale_rows=2,
            source_production_rows=1,
        )

        response = self.client.get(reverse("reportes:cierre_producto"), {"month": "2025-09"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cierre mensual de producto terminado")
        self.assertContains(response, "Detalle del cierre mensual")
        self.assertContains(response, "Meses disponibles")
        self.assertContains(response, "Pastel Cierre Operativo")
        self.assertContains(response, "Producto terminado")
        self.assertContains(response, "bi-closure-table")
        self.assertContains(response, "col-receta-padre")
        self.assertIn("closure", response.context)
        self.assertEqual(response.context["closure"].id, closure.id)
        self.assertEqual(response.context["total_sales"], Decimal("10"))
        self.assertEqual(response.context["total_ending"], Decimal("14"))

    @patch("reportes.views.timezone.localdate", return_value=date(2026, 9, 1))
    def test_cierre_producto_month_options_use_calendar_months_and_include_all_historical_closures(self, _localdate):
        for month_start in (date(2024, 1, 1), date(2025, 12, 1)):
            ProductoMonthClosure.objects.create(
                month_start=month_start,
                month_end=date(month_start.year, month_start.month, 28),
                status=ProductoMonthClosure.STATUS_BUILT,
                opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
            )

        with self.assertNumQueries(1):
            values = [option["value"] for option in _product_closure_month_options(date(2026, 9, 1))]

        self.assertEqual(values[:4], ["2026-09", "2026-08", "2026-07", "2026-06"])
        self.assertIn("2025-12", values)
        self.assertIn("2024-01", values)

    def test_cierre_producto_post_builds_month_if_missing(self):
        admin_user = User.objects.create_user(username="admin_build_cierre", password="pass123")
        admin_group, _ = Group.objects.get_or_create(name="ADMIN")
        admin_user.groups.add(admin_group)
        self.client.force_login(admin_user)

        sucursal = self._create_sucursal("CLOSE-01", "Sucursal Cierre")
        point_branch = PointBranch.objects.create(external_id="CLOSE-01", name="Sucursal Cierre", erp_branch=sucursal)
        sync_job = PointSyncJob.objects.create(
            job_type=PointSyncJob.JOB_TYPE_INVENTORY,
            status=PointSyncJob.STATUS_SUCCESS,
        )
        receta = Receta.objects.create(
            nombre="Pastel Build Cierre",
            codigo_point="CLOSE001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-build-cierre-001",
        )
        point_product = PointProduct.objects.create(
            external_id="CLOSE-PROD-01",
            sku="CLOSE001",
            name=receta.nombre,
            active=True,
        )
        PointInventorySnapshot.objects.create(
            branch=point_branch,
            product=point_product,
            stock=Decimal("9"),
            sync_job=sync_job,
            captured_at=timezone.make_aware(datetime(2025, 8, 31, 23, 0, 0), timezone.get_current_timezone()),
        )

        response = self.client.post(
            reverse("reportes:cierre_producto"),
            {"month": "2025-09"},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Se construyó el cierre calculado de producto Point para 2025-09.")
        self.assertNotContains(response, "cierre teórico")
        closure = ProductoMonthClosure.objects.get(month_start=date(2025, 9, 1))
        self.assertEqual(closure.built_by, admin_user)
        line = closure.lines.get(receta_padre=receta)
        self.assertEqual(line.inventario_inicial_teorico, Decimal("9"))
        self.assertContains(response, "Pastel Build Cierre")
        self.assertContains(response, "9.00")

    def test_cierre_producto_build_rejects_lectura_user(self):
        sucursal = self._create_sucursal("CLOSE-02", "Sucursal Cierre 02")
        point_branch = PointBranch.objects.create(external_id="CLOSE-02", name="Sucursal Cierre 02", erp_branch=sucursal)
        sync_job = PointSyncJob.objects.create(
            job_type=PointSyncJob.JOB_TYPE_INVENTORY,
            status=PointSyncJob.STATUS_SUCCESS,
        )
        receta = Receta.objects.create(
            nombre="Pastel Build Prohibido",
            codigo_point="CLOSE002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-build-prohibido-001",
        )
        point_product = PointProduct.objects.create(
            external_id="CLOSE-PROD-02",
            sku="CLOSE002",
            name=receta.nombre,
            active=True,
        )
        PointInventorySnapshot.objects.create(
            branch=point_branch,
            product=point_product,
            stock=Decimal("6"),
            sync_job=sync_job,
            captured_at=timezone.make_aware(datetime(2025, 8, 31, 23, 0, 0), timezone.get_current_timezone()),
        )

        response = self.client.post(
            reverse("reportes:cierre_producto"),
            {"month": "2025-09"},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No tienes permisos para construir cierres mensuales de producto.")
        self.assertFalse(ProductoMonthClosure.objects.filter(month_start=date(2025, 9, 1), built_by=self.user).exists())

    def test_cierre_producto_export_csv_returns_file(self):
        receta = Receta.objects.create(
            nombre="Pastel Export Cierre",
            codigo_point="EXP001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-export-cierre-001",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2025, 12, 1),
            month_end=date(2025, 12, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
            opening_reference_date=date(2025, 11, 30),
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=receta,
            inventario_inicial_teorico=Decimal("11"),
            produccion_mes=Decimal("2"),
            venta_total_equivalente=Decimal("4"),
            merma_total_equivalente=Decimal("1"),
            inventario_final_teorico=Decimal("8"),
        )

        response = self.client.get(reverse("reportes:cierre_producto"), {"month": "2025-12", "export": "csv"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        self.assertIn("attachment; filename=\"cierre_producto_2025-12.csv\"", response["Content-Disposition"])
        self.assertIn("Pastel Export Cierre", response.content.decode("utf-8"))

    def test_cierre_producto_exports_use_canonical_point_sign_labels_and_status(self):
        receta = Receta.objects.create(
            nombre="Pastel Export Canonico",
            codigo_point="EXPCAN001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-export-canonico-001",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2026, 8, 1),
            month_end=date(2026, 8, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
            opening_reference_date=date(2026, 7, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=receta,
            inventario_inicial_teorico=Decimal("10"),
            inventario_final_teorico=Decimal("9"),
            inventario_final_point_total=Decimal("11"),
            diferencia_teorico_vs_point=Decimal("-2"),
            metadata={
                "balance_contract": "POINT_PRODUCT_BALANCE_V1",
                "point_difference": "2",
                "point_status": "POINT_MAYOR",
                "issues": [],
                "sales_source_available": True,
                "production_source_authoritative": True,
                "waste_source_authoritative": True,
                "conversion_source_authoritative": True,
                "point_conversion_in": "8",
                "point_conversion_out": "1",
                "conversion_origin": "MIXED",
                "conversion_origins": ["EQUIVALENCIA_CONFIGURADA", "POINT"],
                "projection_sources": ["EQUIVALENCIA"],
            },
        )

        html_response = self.client.get(reverse("reportes:cierre_producto"), {"month": "2026-08"})
        html_body = html_response.content.decode("utf-8")
        self.assertIn("Origen conversión", html_body)
        self.assertIn("Proyección", html_body)
        self.assertIn("EQUIVALENCIA_CONFIGURADA, POINT", html_body)
        self.assertIn(">EQUIVALENCIA<", html_body)

        csv_response = self.client.get(
            reverse("reportes:cierre_producto"),
            {"month": "2026-08", "export": "csv"},
        )
        csv_body = csv_response.content.decode("utf-8")
        self.assertIn("Saldo calculado", csv_body)
        self.assertIn("Fin. Point", csv_body)
        self.assertIn("Dif. Point", csv_body)
        csv_rows = list(csv.reader(StringIO(csv_body)))
        detail_headers = csv_rows[csv_rows.index([]) + 1]
        detail_values = csv_rows[csv_rows.index([]) + 2]
        self.assertEqual(Decimal(detail_values[detail_headers.index("Saldo calculado")]), Decimal("9"))
        self.assertEqual(Decimal(detail_values[detail_headers.index("Fin. Point")]), Decimal("11"))
        self.assertEqual(Decimal(detail_values[detail_headers.index("Dif. Point")]), Decimal("2"))
        self.assertEqual(Decimal(detail_values[detail_headers.index("Venta directa")]), Decimal("0"))
        self.assertEqual(Decimal(detail_values[detail_headers.index("Venta derivada")]), Decimal("0"))
        self.assertEqual(Decimal(detail_values[detail_headers.index("Conv. entrada Point")]), Decimal("8"))
        self.assertEqual(Decimal(detail_values[detail_headers.index("Conv. salida Point")]), Decimal("1"))
        self.assertEqual(
            detail_values[detail_headers.index("Origen conversión")],
            "EQUIVALENCIA_CONFIGURADA | POINT",
        )
        self.assertEqual(detail_values[detail_headers.index("Proyección")], "EQUIVALENCIA")
        self.assertEqual(detail_values[detail_headers.index("Estado")], "Point mayor")
        self.assertNotIn("Diferencia teorico vs Point", csv_body)
        self.assertNotIn("Sobrante físico", csv_body)
        self.assertNotIn("Faltante no explicado", csv_body)

        xlsx_response = self.client.get(
            reverse("reportes:cierre_producto"),
            {"month": "2026-08", "export": "xlsx"},
        )
        workbook = load_workbook(BytesIO(xlsx_response.content), data_only=True, read_only=True)
        detail = workbook["Detalle"]
        headers = [cell.value for cell in next(detail.iter_rows(min_row=1, max_row=1))]
        values = next(detail.iter_rows(min_row=2, max_row=2, values_only=True))
        self.assertIn("Saldo calculado", headers)
        self.assertIn("Fin. Point", headers)
        self.assertIn("Dif. Point", headers)
        self.assertEqual(values[headers.index("Saldo calculado")], 9)
        self.assertEqual(values[headers.index("Fin. Point")], 11)
        self.assertEqual(values[headers.index("Dif. Point")], 2)
        self.assertEqual(values[headers.index("Venta directa")], 0)
        self.assertEqual(values[headers.index("Venta derivada")], 0)
        self.assertEqual(values[headers.index("Conv. entrada Point")], 8)
        self.assertEqual(values[headers.index("Conv. salida Point")], 1)
        self.assertEqual(values[headers.index("Origen conversión")], "EQUIVALENCIA_CONFIGURADA | POINT")
        self.assertEqual(values[headers.index("Proyección")], "EQUIVALENCIA")
        self.assertEqual(values[headers.index("Estado")], "Point mayor")

    def test_historical_mixed_conversion_metadata_is_split_conservatively_in_outputs(self):
        receta = Receta.objects.create(
            nombre="Pastel trazabilidad histórica",
            codigo_point="HIST-TRACE-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-historical-conversion-trace",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2026, 4, 1),
            month_end=date(2026, 4, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
        )
        line = ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=receta,
            metadata={
                "balance_contract": "POINT_PRODUCT_BALANCE_V1",
                "point_conversion_in": "4",
                "point_conversion_out": "1",
                "conversion_origin": ["POINT", "PRESENTACION_DERIVADA"],
                "conversion_source_authoritative": True,
                "production_source_authoritative": True,
                "waste_source_authoritative": True,
                "sales_source_available": True,
                "issues": [],
            },
        )

        response = self.client.get(reverse("reportes:cierre_producto"), {"month": "2026-04"})
        conversion = response.context["conversion_rows"][0]
        self.assertEqual(conversion["conversion_origin"], ("POINT",))
        self.assertEqual(conversion["projection_sources"], ("PRESENTACION_DERIVADA",))

        csv_response = self.client.get(
            reverse("reportes:cierre_producto"), {"month": "2026-04", "export": "csv"}
        )
        rows = list(csv.reader(StringIO(csv_response.content.decode("utf-8"))))
        headers = rows[rows.index([]) + 1]
        values = rows[rows.index([]) + 2]
        self.assertEqual(values[headers.index("Origen conversión")], "POINT")
        self.assertEqual(values[headers.index("Proyección")], "PRESENTACION_DERIVADA")

        xlsx_response = self.client.get(
            reverse("reportes:cierre_producto"), {"month": "2026-04", "export": "xlsx"}
        )
        workbook = load_workbook(BytesIO(xlsx_response.content), data_only=True, read_only=True)
        headers = [cell.value for cell in next(workbook["Detalle"].iter_rows(min_row=1, max_row=1))]
        values = next(workbook["Detalle"].iter_rows(min_row=2, max_row=2, values_only=True))
        self.assertEqual(values[headers.index("Origen conversión")], "POINT")
        self.assertEqual(values[headers.index("Proyección")], "PRESENTACION_DERIVADA")

        line.metadata = {
            key: value
            for key, value in line.metadata.items()
            if key != "conversion_origins"
        }
        line.metadata["conversion_origin"] = "MIXED"
        line.save(update_fields=["metadata", "updated_at"])
        scalar_response = self.client.get(reverse("reportes:cierre_producto"), {"month": "2026-04"})
        self.assertEqual(scalar_response.context["conversion_rows"][0]["conversion_origin"], ("Mixto",))

    def test_unresolved_conversion_origin_is_visible_without_inventing_components(self):
        receta = Receta.objects.create(
            nombre="Conversión sin origen resuelto",
            codigo_point="UNRESOLVED-CONV-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-unresolved-conversion-output",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=receta,
            metadata={
                "balance_contract": "POINT_PRODUCT_BALANCE_V1",
                "point_conversion_in": "1",
                "point_conversion_out": "0",
                "conversion_origin": "UNRESOLVED",
                "conversion_origins": ["UNRESOLVED"],
                "conversion_source_authoritative": True,
                "production_source_authoritative": True,
                "waste_source_authoritative": True,
                "sales_source_available": True,
                "issues": ["CONVERSION_ORIGIN_UNRESOLVED"],
            },
        )

        response = self.client.get(reverse("reportes:cierre_producto"), {"month": "2026-03"})
        conversion = response.context["conversion_rows"][0]
        self.assertEqual(conversion["conversion_origin"], ("Sin resolver",))
        self.assertNotIn("POINT", conversion["conversion_origin"])
        self.assertNotIn("EQUIVALENCIA_CONFIGURADA", conversion["conversion_origin"])

    def test_cierre_producto_html_uses_neutral_point_balance_and_canonical_sign(self):
        receta = Receta.objects.create(
            nombre="Pastel HTML Canonico",
            codigo_point="HTMLCAN001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-html-canonico-001",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2026, 8, 1),
            month_end=date(2026, 8, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
            opening_reference_date=date(2026, 7, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=receta,
            inventario_inicial_teorico=Decimal("10"),
            inventario_final_teorico=Decimal("9"),
            inventario_final_point_total=Decimal("11"),
            diferencia_teorico_vs_point=Decimal("-2"),
            estado_auditoria=ProductoMonthClosureLine.AUDIT_STATUS_SOBRANTE_FISICO,
            metadata={
                "balance_contract": "POINT_PRODUCT_BALANCE_V1",
                "point_difference": "2",
                "point_status": "POINT_MAYOR",
                "issues": [],
            },
        )

        response = self.client.get(reverse("reportes:cierre_producto"), {"month": "2026-08"})
        body = response.content.decode("utf-8")

        self.assertContains(response, "Saldo calculado")
        self.assertContains(response, "Fin. Point")
        self.assertContains(response, "Dif. Point")
        self.assertContains(response, "Point mayor")
        self.assertEqual(response.context["total_closing_difference"], Decimal("2"))
        self.assertNotIn("Point físico", body)
        self.assertNotIn("Sobrante físico", body)
        self.assertNotIn("Faltante no explicado", body)
        self.assertNotIn("Final teórico", body)
        self.assertNotIn("final teórico", body)
        self.assertNotIn("cierre teórico", body.lower())

    def test_historical_non_point_opening_uses_honest_labels_in_html_csv_and_xlsx(self):
        receta = Receta.objects.create(
            nombre="Pastel Apertura Historica",
            codigo_point="HISTOPEN001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-apertura-historica-001",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2025, 10, 1),
            month_end=date(2025, 10, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_PREVIOUS_CLOSURE,
            opening_reference_date=date(2025, 9, 30),
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=receta,
            inventario_inicial_teorico=Decimal("7"),
            inventario_final_teorico=Decimal("7"),
            inventario_final_point_total=Decimal("7"),
            diferencia_teorico_vs_point=Decimal("0"),
        )

        html_response = self.client.get(reverse("reportes:cierre_producto"), {"month": "2025-10"})
        self.assertContains(html_response, "Saldo inicial")
        self.assertContains(html_response, "Cierre previo")

        csv_response = self.client.get(
            reverse("reportes:cierre_producto"), {"month": "2025-10", "export": "csv"}
        )
        csv_rows = list(csv.reader(StringIO(csv_response.content.decode("utf-8"))))
        summary = dict(row for row in csv_rows[: csv_rows.index([])] if len(row) == 2)
        detail_headers = csv_rows[csv_rows.index([]) + 1]
        self.assertEqual(summary["Fuente inicial"], "Cierre previo")
        self.assertEqual(Decimal(summary["Saldo inicial"]), Decimal("7"))
        self.assertIn("Saldo inicial", detail_headers)
        self.assertNotIn("Ini. Point", detail_headers)

        xlsx_response = self.client.get(
            reverse("reportes:cierre_producto"), {"month": "2025-10", "export": "xlsx"}
        )
        workbook = load_workbook(BytesIO(xlsx_response.content), data_only=True, read_only=True)
        summary_values = {
            row[0]: row[1]
            for row in workbook["Resumen"].iter_rows(min_col=1, max_col=2, values_only=True)
            if row[0]
        }
        headers = [cell.value for cell in next(workbook["Detalle"].iter_rows(min_row=1, max_row=1))]
        self.assertEqual(summary_values["Fuente inicial"], "Cierre previo")
        self.assertEqual(summary_values["Saldo inicial"], 7)
        self.assertIn("Saldo inicial", headers)
        self.assertNotIn("Ini. Point", headers)

    def test_cierre_producto_exports_do_not_turn_missing_canonical_sources_into_zero(self):
        receta = Receta.objects.create(
            nombre="Pastel Export Sin Fuente",
            codigo_point="EXPMISS001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-export-sin-fuente-001",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2026, 7, 1),
            month_end=date(2026, 7, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
            opening_reference_date=date(2026, 6, 30),
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=receta,
            inventario_final_teorico=Decimal("0"),
            inventario_final_point_total=Decimal("0"),
            diferencia_teorico_vs_point=Decimal("0"),
            metadata={
                "balance_contract": "POINT_PRODUCT_BALANCE_V1",
                "point_difference": "",
                "point_status": "REVISAR_FUENTE",
                "issues": [
                    "OPENING_SNAPSHOT_MISSING",
                    "CALCULATED_CLOSING_MISSING",
                    "CLOSING_SNAPSHOT_MISSING",
                ],
            },
        )

        csv_response = self.client.get(
            reverse("reportes:cierre_producto"),
            {"month": "2026-07", "export": "csv"},
        )
        csv_body = csv_response.content.decode("utf-8")
        csv_rows = list(csv.reader(StringIO(csv_body)))
        detail_headers = csv_rows[csv_rows.index([]) + 1]
        detail_values = csv_rows[csv_rows.index([]) + 2]
        summary = dict(row for row in csv_rows[: csv_rows.index([])] if len(row) == 2)
        self.assertEqual(summary["Fuente Ini. Point"], "Snapshot Point")
        self.assertEqual(summary["Ini. Point"], "Sin dato")
        self.assertIn("Ini. Point", detail_headers)
        self.assertEqual(detail_values[detail_headers.index("Ini. Point")], "Sin dato")
        self.assertEqual(detail_values[detail_headers.index("Saldo calculado")], "Sin dato")
        self.assertEqual(detail_values[detail_headers.index("Produccion")], "Sin dato")
        self.assertEqual(detail_values[detail_headers.index("Merma total")], "Sin dato")
        self.assertEqual(detail_values[detail_headers.index("Conv. entrada Point")], "Sin dato")
        self.assertEqual(detail_values[detail_headers.index("Conv. salida Point")], "Sin dato")
        self.assertEqual(detail_values[detail_headers.index("Fin. Point")], "Sin dato")
        self.assertEqual(detail_values[detail_headers.index("Dif. Point")], "Sin dato")
        self.assertEqual(detail_values[detail_headers.index("Estado")], "Revisar fuente")

        html_response = self.client.get(reverse("reportes:cierre_producto"), {"month": "2026-07"})
        self.assertIsNone(html_response.context["total_opening"])
        self.assertIsNone(html_response.context["total_ending"])
        self.assertIsNone(html_response.context["total_closing_point"])
        self.assertIsNone(html_response.context["total_closing_difference"])
        self.assertContains(html_response, "Sin dato")

        xlsx_response = self.client.get(
            reverse("reportes:cierre_producto"),
            {"month": "2026-07", "export": "xlsx"},
        )
        workbook = load_workbook(BytesIO(xlsx_response.content), data_only=True, read_only=True)
        summary_sheet = workbook["Resumen"]
        summary_values = {
            row[0]: row[1]
            for row in summary_sheet.iter_rows(min_col=1, max_col=2, values_only=True)
            if row[0]
        }
        detail = workbook["Detalle"]
        headers = [cell.value for cell in next(detail.iter_rows(min_row=1, max_row=1))]
        values = next(detail.iter_rows(min_row=2, max_row=2, values_only=True))
        self.assertEqual(summary_values["Ini. Point"], "Sin dato")
        self.assertEqual(summary_values["Fuente Ini. Point"], "Snapshot Point")
        self.assertIn("Ini. Point", headers)
        self.assertEqual(values[headers.index("Ini. Point")], "Sin dato")
        self.assertEqual(values[headers.index("Saldo calculado")], "Sin dato")
        self.assertEqual(values[headers.index("Produccion")], "Sin dato")
        self.assertEqual(values[headers.index("Merma total")], "Sin dato")
        self.assertEqual(values[headers.index("Conv. entrada Point")], "Sin dato")
        self.assertEqual(values[headers.index("Conv. salida Point")], "Sin dato")
        self.assertEqual(values[headers.index("Fin. Point")], "Sin dato")
        self.assertEqual(values[headers.index("Dif. Point")], "Sin dato")
        self.assertEqual(values[headers.index("Estado")], "Revisar fuente")

    def test_real_missing_sales_propagates_sin_dato_through_closure_and_recent_card(self):
        admin_user = User.objects.create_user(username="admin_missing_sales", password="pass123")
        admin_group, _ = Group.objects.get_or_create(name="ADMIN")
        admin_user.groups.add(admin_group)
        self.client.force_login(admin_user)
        sucursal = self._create_sucursal("NO-SALES", "Sucursal sin ventas")
        branch = PointBranch.objects.create(external_id="NO-SALES", name="Sucursal sin ventas", erp_branch=sucursal)
        job = PointSyncJob.objects.create(
            job_type=PointSyncJob.JOB_TYPE_INVENTORY,
            status=PointSyncJob.STATUS_SUCCESS,
        )
        receta = Receta.objects.create(
            nombre="Pastel sin fuente venta",
            codigo_point="NO-SALES-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-no-sales-001",
        )
        product = PointProduct.objects.create(
            external_id="NO-SALES-PRODUCT",
            sku=receta.codigo_point,
            name=receta.nombre,
            active=True,
        )
        for captured_at in (datetime(2026, 6, 30, 8), datetime(2026, 7, 31, 8)):
            PointInventorySnapshot.objects.create(
                branch=branch,
                product=product,
                stock=Decimal("5"),
                captured_at=timezone.make_aware(captured_at, timezone.get_current_timezone()),
                sync_job=job,
            )

        response = self.client.post(
            reverse("reportes:cierre_producto"),
            {"month": "2026-07", "action": "build"},
            follow=True,
        )
        closure = ProductoMonthClosure.objects.get(month_start=date(2026, 7, 1))
        line = closure.lines.get(receta_padre=receta)

        self.assertIn("CALCULATED_CLOSING_MISSING", line.metadata["issues"])
        self.assertIn("SALES_SOURCE_MISSING", line.metadata["issues"])
        self.assertIsNone(response.context["closure_display_rows"][0]["sales_total"])
        self.assertIsNone(response.context["total_sales"])
        self.assertIsNone(response.context["total_ending"])
        self.assertIsNone(response.context["recent_closure_rows"][0]["ending_inventory"])
        self.assertIsNone(response.context["recent_closure_rows"][0]["sales_total"])
        self.assertContains(response, "Sin dato")

        csv_response = self.client.get(
            reverse("reportes:cierre_producto"), {"month": "2026-07", "export": "csv"}
        )
        csv_rows = list(csv.reader(StringIO(csv_response.content.decode("utf-8"))))
        headers = csv_rows[csv_rows.index([]) + 1]
        values = csv_rows[csv_rows.index([]) + 2]
        summary = dict(row for row in csv_rows[: csv_rows.index([])] if len(row) == 2)
        self.assertEqual(summary["Venta equivalente"], "Sin dato")
        self.assertEqual(values[headers.index("Venta directa")], "Sin dato")
        self.assertEqual(values[headers.index("Venta derivada")], "Sin dato")
        self.assertEqual(values[headers.index("Saldo calculado")], "Sin dato")
        self.assertEqual(values[headers.index("Dif. Point")], "Sin dato")

        xlsx_response = self.client.get(
            reverse("reportes:cierre_producto"), {"month": "2026-07", "export": "xlsx"}
        )
        workbook = load_workbook(BytesIO(xlsx_response.content), data_only=True, read_only=True)
        xlsx_headers = [cell.value for cell in next(workbook["Detalle"].iter_rows(min_row=1, max_row=1))]
        xlsx_values = next(workbook["Detalle"].iter_rows(min_row=2, max_row=2, values_only=True))
        xlsx_summary = {
            row[0]: row[1]
            for row in workbook["Resumen"].iter_rows(min_col=1, max_col=2, values_only=True)
            if row[0]
        }
        self.assertEqual(xlsx_summary["Venta equivalente"], "Sin dato")
        self.assertEqual(xlsx_values[xlsx_headers.index("Venta directa")], "Sin dato")
        self.assertEqual(xlsx_values[xlsx_headers.index("Venta derivada")], "Sin dato")
        self.assertEqual(xlsx_values[xlsx_headers.index("Saldo calculado")], "Sin dato")
        self.assertEqual(xlsx_values[xlsx_headers.index("Dif. Point")], "Sin dato")

    def test_secondary_closure_panels_project_missing_sales_and_opening_as_sin_dato(self):
        receta = Receta.objects.create(
            nombre="Pastel panel sin fuentes",
            codigo_point="PANEL-MISSING-001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-panel-missing-001",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2026, 6, 1),
            month_end=date(2026, 6, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=receta,
            venta_derivada_equivalente=Decimal("9"),
            venta_total_equivalente=Decimal("9"),
            merma_derivada_equivalente=Decimal("1"),
            merma_total_equivalente=Decimal("1"),
            inventario_inicial_teorico=Decimal("0"),
            inventario_final_teorico=Decimal("0"),
            has_catalog_issue=True,
            catalog_issue_note="Fuentes canónicas ausentes",
            metadata={
                "balance_contract": "POINT_PRODUCT_BALANCE_V1",
                "issues": [
                    "SALES_SOURCE_MISSING",
                    "OPENING_SNAPSHOT_MISSING",
                    "CALCULATED_CLOSING_MISSING",
                ],
                "sales_source_available": False,
                "point_status": "REVISAR_FUENTE",
            },
        )

        response = self.client.get(reverse("reportes:cierre_producto"), {"month": "2026-06"})

        catalog_row = response.context["catalog_issue_rows"][0]
        self.assertEqual(response.context["conversion_rows"], [])
        self.assertIsNone(catalog_row["calculated_closing"])
        body = response.content.decode("utf-8")
        conversion_panel = body.split("Conversión Point registrada", 1)[1].split("Riesgo de catálogo", 1)[0]
        catalog_panel = body.split("Incidencias visibles", 1)[1].split("Auditoría de líneas", 1)[0]
        self.assertIn("No hubo conversiones Point registradas", conversion_panel)
        self.assertNotIn(">9.00<", conversion_panel)
        self.assertNotIn(">0.00<", conversion_panel)
        self.assertIn("Sin dato", catalog_panel)
        self.assertNotIn(">0.00<", catalog_panel)

    def test_conversion_panel_uses_only_canonical_point_conversion_movements(self):
        sale_parent = Receta.objects.create(
            nombre="Pastel con venta derivada sin conversión Point",
            codigo_point="NO-CONV-POINT",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-no-conv-point",
        )
        converted_parent = Receta.objects.create(
            nombre="Pastel con conversión Point",
            codigo_point="WITH-CONV-POINT",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-with-conv-point",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2026, 5, 1),
            month_end=date(2026, 5, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=sale_parent,
            venta_derivada_equivalente=Decimal("8"),
            merma_derivada_equivalente=Decimal("1"),
            metadata={
                "balance_contract": "POINT_PRODUCT_BALANCE_V1",
                "point_conversion_in": "0",
                "point_conversion_out": "0",
                "conversion_origin": [],
                "conversion_source_authoritative": True,
                "issues": [],
                "sales_source_available": True,
            },
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=converted_parent,
            metadata={
                "balance_contract": "POINT_PRODUCT_BALANCE_V1",
                "point_conversion_in": "16",
                "point_conversion_out": "2",
                "conversion_origin": ["EQUIVALENCIA_CONFIGURADA"],
                "conversion_source_authoritative": True,
                "issues": [],
                "sales_source_available": True,
            },
        )

        response = self.client.get(reverse("reportes:cierre_producto"), {"month": "2026-05"})

        self.assertEqual(len(response.context["conversion_rows"]), 1)
        conversion = response.context["conversion_rows"][0]
        self.assertEqual(conversion["line"].receta_padre_id, converted_parent.id)
        self.assertEqual(conversion["point_conversion_in"], Decimal("16"))
        self.assertEqual(conversion["point_conversion_out"], Decimal("2"))
        self.assertEqual(conversion["conversion_origin"], ("EQUIVALENCIA_CONFIGURADA",))
        panel = response.content.decode("utf-8").split("Conversión Point registrada", 1)[1].split(
            "Riesgo de catálogo", 1
        )[0]
        self.assertIn("Pastel con conversión Point", panel)
        self.assertNotIn("Pastel con venta derivada sin conversión Point", panel)

    def test_historical_excel_inventory_is_not_presented_as_point_and_canonical_remains_point(self):
        receta = Receta.objects.create(
            nombre="Pastel Export Historico",
            codigo_point="EXPHIST001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-export-historico-001",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2025, 6, 1),
            month_end=date(2025, 6, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_BOOTSTRAP_SEED,
            opening_reference_date=date(2025, 5, 31),
            metadata={"historical_excel_import": {"scope": "inventory_only", "source_file": "junio.xlsx"}},
        )
        line = ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=receta,
            inventario_final_teorico=Decimal("7"),
            inventario_final_point_total=Decimal("10"),
            diferencia_teorico_vs_point=Decimal("-3"),
            estado_auditoria=ProductoMonthClosureLine.AUDIT_STATUS_SOBRANTE_FISICO,
            metadata={"historical_excel": {"scope": "inventory_only"}},
        )

        html_response = self.client.get(reverse("reportes:cierre_producto"), {"month": "2025-06"})
        html_body = html_response.content.decode("utf-8")
        self.assertIn("Inventario histórico físico", html_body)
        self.assertIn("Diferencia histórica", html_body)
        self.assertIn("Sobrante físico", html_body)
        self.assertIn("Importación histórica de Excel", html_body)
        self.assertNotIn("Fin. Point", html_body)
        self.assertNotIn("Dif. Point", html_body)

        response = self.client.get(
            reverse("reportes:cierre_producto"),
            {"month": "2025-06", "export": "csv"},
        )
        rows = list(csv.reader(StringIO(response.content.decode("utf-8"))))
        headers = rows[rows.index([]) + 1]
        values = rows[rows.index([]) + 2]

        self.assertIn("Inventario histórico físico", headers)
        self.assertIn("Diferencia histórica", headers)
        self.assertNotIn("Fin. Point", headers)
        self.assertNotIn("Dif. Point", headers)
        self.assertEqual(Decimal(values[headers.index("Inventario histórico físico")]), Decimal("10"))
        self.assertEqual(Decimal(values[headers.index("Diferencia histórica")]), Decimal("-3"))
        self.assertEqual(values[headers.index("Estado")], "Sobrante físico")

        xlsx_response = self.client.get(
            reverse("reportes:cierre_producto"),
            {"month": "2025-06", "export": "xlsx"},
        )
        workbook = load_workbook(BytesIO(xlsx_response.content), data_only=True, read_only=True)
        xlsx_headers = [cell.value for cell in next(workbook["Detalle"].iter_rows(min_row=1, max_row=1))]
        xlsx_values = next(workbook["Detalle"].iter_rows(min_row=2, max_row=2, values_only=True))
        self.assertIn("Inventario histórico físico", xlsx_headers)
        self.assertIn("Diferencia histórica", xlsx_headers)
        self.assertNotIn("Fin. Point", xlsx_headers)
        self.assertEqual(xlsx_values[xlsx_headers.index("Diferencia histórica")], -3)

        canonical = ProductoMonthClosure.objects.create(
            month_start=date(2026, 6, 1),
            month_end=date(2026, 6, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
            metadata={"balance": {"contract": "POINT_PRODUCT_BALANCE_V1"}},
        )
        ProductoMonthClosureLine.objects.create(
            closure=canonical,
            receta_padre=receta,
            inventario_final_teorico=Decimal("7"),
            inventario_final_point_total=Decimal("10"),
            diferencia_teorico_vs_point=Decimal("-3"),
            metadata={
                "balance_contract": "POINT_PRODUCT_BALANCE_V1",
                "point_difference": "3",
                "point_status": "POINT_MAYOR",
                "issues": [],
            },
        )
        canonical_csv = self.client.get(
            reverse("reportes:cierre_producto"),
            {"month": "2026-06", "export": "csv"},
        ).content.decode("utf-8")
        self.assertIn("Fin. Point", canonical_csv)
        self.assertIn("Dif. Point", canonical_csv)
        self.assertNotIn("Inventario histórico físico", canonical_csv)
        line.refresh_from_db()
        self.assertEqual(line.diferencia_teorico_vs_point, Decimal("-3"))
        self.assertEqual(line.estado_auditoria, ProductoMonthClosureLine.AUDIT_STATUS_SOBRANTE_FISICO)

    def test_cierre_producto_export_keeps_total_when_only_point_scope_split_is_missing(self):
        receta = Receta.objects.create(
            nombre="Pastel Export Scope Parcial",
            codigo_point="EXPSCOPE001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-export-scope-parcial-001",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2026, 5, 1),
            month_end=date(2026, 5, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
            opening_reference_date=date(2026, 4, 30),
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=receta,
            inventario_final_teorico=Decimal("8"),
            inventario_final_point_total=Decimal("10"),
            diferencia_teorico_vs_point=Decimal("-2"),
            metadata={
                "balance_contract": "POINT_PRODUCT_BALANCE_V1",
                "point_difference": "2",
                "point_status": "REVISAR_FUENTE",
                "point_final_scopes_available": False,
                "issues": ["CLOSING_SNAPSHOT_SCOPE_MISSING"],
            },
        )

        response = self.client.get(
            reverse("reportes:cierre_producto"),
            {"month": "2026-05", "export": "csv"},
        )
        rows = list(csv.reader(StringIO(response.content.decode("utf-8"))))
        headers = rows[rows.index([]) + 1]
        values = rows[rows.index([]) + 2]

        self.assertEqual(values[headers.index("Point CEDIS")], "Sin dato")
        self.assertEqual(values[headers.index("Point sucursales")], "Sin dato")
        self.assertEqual(Decimal(values[headers.index("Fin. Point")]), Decimal("10"))
        self.assertEqual(Decimal(values[headers.index("Dif. Point")]), Decimal("2"))
        self.assertEqual(values[headers.index("Estado")], "Revisar fuente")

    def test_cierre_producto_lock_requires_admin_or_dg(self):
        receta = Receta.objects.create(
            nombre="Pastel Lock Lectura",
            codigo_point="LOCK001",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-lock-lectura-001",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2025, 9, 1),
            month_end=date(2025, 9, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
            opening_reference_date=date(2025, 8, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=receta,
            inventario_inicial_teorico=Decimal("4"),
            inventario_final_teorico=Decimal("4"),
        )

        response = self.client.post(
            reverse("reportes:cierre_producto"),
            {"month": "2025-09", "action": "lock"},
            follow=True,
        )

        closure.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertFalse(closure.is_locked)
        self.assertContains(response, "Solo Dirección General o Administración pueden bloquear cierres mensuales.")

    def test_cierre_producto_disables_lock_and_explains_each_movement_authority_guard(self):
        admin_user = User.objects.create_user(username="admin_cierre_fuentes", password="pass123")
        admin_group, _ = Group.objects.get_or_create(name="ADMIN")
        admin_user.groups.add(admin_group)
        self.client.force_login(admin_user)
        cases = (
            (
                "production",
                date(2026, 1, 1),
                ("PRODUCTION_SYNC_JOB_PARTIAL", "PRODUCTION_SYNC_RANGE_INCOMPLETE"),
                ("Producción: job Point parcial", "Producción: rango mensual incompleto"),
            ),
            (
                "waste",
                date(2026, 2, 1),
                ("WASTE_SYNC_JOB_MISSING", "WASTE_SYNC_JOB_RESTRICTED"),
                ("Merma: falta job Point del mes", "Merma: job Point filtrado por sucursal"),
            ),
            (
                "conversion",
                date(2026, 3, 1),
                (
                    "CONVERSION_SYNC_CONTRACT_INCOMPLETE",
                    "CONVERSION_SYNC_COUNT_MISMATCH",
                    "CONVERSION_SYNC_JOB_MIXED",
                ),
                (
                    "Conversiones: contrato del job incompleto",
                    "Conversiones: conteo del job no reconcilia",
                    "Conversiones: filas mezcladas entre jobs",
                ),
            ),
        )
        for family, month_start, authority_issues, expected_messages in cases:
            with self.subTest(family=family):
                receta = Receta.objects.create(
                    nombre=f"Pastel guarda {family}",
                    codigo_point=f"GUARD-{family.upper()}",
                    tipo=Receta.TIPO_PRODUCTO_FINAL,
                    hash_contenido=f"hash-guard-{family}",
                )
                metadata_key = f"{family}_meta"
                closure = ProductoMonthClosure.objects.create(
                    month_start=month_start,
                    month_end=date(month_start.year, month_start.month, 28),
                    status=ProductoMonthClosure.STATUS_BUILT,
                    opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
                    metadata={
                        "validation": {
                            "lock_ready": False,
                            "blocking_issues": ["MONTH_SOURCE_INCOMPLETE"],
                        },
                        metadata_key: {
                            "source": {
                                "production": "PointProductionLine",
                                "waste": "PointWasteLine",
                                "conversion": "PointConversionLine",
                            }[family],
                            "authoritative": False,
                            "source_present": family != "waste",
                            "job_status": "PARTIAL" if family == "production" else "SUCCESS",
                            "authority_issues": list(authority_issues),
                        },
                    },
                )
                ProductoMonthClosureLine.objects.create(
                    closure=closure,
                    receta_padre=receta,
                    metadata={
                        "balance_contract": "POINT_PRODUCT_BALANCE_V1",
                        "issues": ["MONTH_SOURCE_INCOMPLETE"],
                    },
                )

                response = self.client.get(
                    reverse("reportes:cierre_producto"),
                    {"month": month_start.strftime("%Y-%m")},
                )

                self.assertTrue(response.context["can_lock_closure"])
                button = re.search(r"<button[^>]*>Bloquear cierre</button>", response.content.decode("utf-8"))
                self.assertIsNotNone(button)
                self.assertIn("disabled", button.group(0))
                self.assertIn(
                    "Guarda de validación: MONTH_SOURCE_INCOMPLETE",
                    response.context["lock_guard_errors"],
                )
                self.assertContains(response, "Guarda de validación: MONTH_SOURCE_INCOMPLETE")
                for message in expected_messages:
                    self.assertIn(message, response.context["lock_guard_errors"])
                    self.assertContains(response, message)
                details = " ".join(row["detail"] for row in response.context["exception_rows"])
                for message in expected_messages:
                    self.assertIn(message, details)

    def test_cierre_producto_sales_authority_guard_is_human_readable_and_disables_lock(self):
        admin_user = User.objects.create_user(username="admin_cierre_ventas", password="pass123")
        admin_group, _ = Group.objects.get_or_create(name="ADMIN")
        admin_user.groups.add(admin_group)
        self.client.force_login(admin_user)
        receta = Receta.objects.create(
            nombre="Pastel guarda ventas",
            codigo_point="GUARD-SALES",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-guard-sales",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2026, 4, 1),
            month_end=date(2026, 4, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
            metadata={
                "validation": {"lock_ready": False, "blocking_issues": []},
                "sales_meta": {
                    "source": "PointDailySale",
                    "authoritative": False,
                    "source_present": True,
                    "authority_issues": [
                        "SALES_SYNC_JOB_MISSING",
                        "OFFICIAL_SALES_REFRESH_REQUIRED",
                    ],
                },
            },
        )
        ProductoMonthClosureLine.objects.create(closure=closure, receta_padre=receta)

        response = self.client.get(reverse("reportes:cierre_producto"), {"month": "2026-04"})

        expected = (
            "Ventas: falta job Point del mes",
            "Ventas: falta actualizar el reporte oficial de Point",
        )
        button = re.search(r"<button[^>]*>Bloquear cierre</button>", response.content.decode("utf-8"))
        self.assertIsNotNone(button)
        self.assertIn("disabled", button.group(0))
        for message in expected:
            self.assertIn(message, response.context["lock_guard_errors"])
            self.assertContains(response, message)
        rendered_exceptions = " ".join(row["detail"] for row in response.context["exception_rows"])
        self.assertNotIn("SALES_SYNC_JOB_MISSING", rendered_exceptions)

        post_response = self.client.post(
            reverse("reportes:cierre_producto"),
            {"month": "2026-04", "action": "lock"},
            follow=True,
        )

        closure.refresh_from_db()
        self.assertFalse(closure.is_locked)
        self.assertContains(post_response, "no esta listo para bloquearse")
        self.assertNotIn("OFFICIAL_SALES_REFRESH_REQUIRED", rendered_exceptions)

    def test_cierre_producto_lock_succeeds_for_admin_when_closure_is_clean(self):
        admin_user = User.objects.create_user(username="admin_cierre", password="pass123")
        admin_group, _ = Group.objects.get_or_create(name="ADMIN")
        admin_user.groups.add(admin_group)
        self.client.force_login(admin_user)

        receta = Receta.objects.create(
            nombre="Pastel Lock Admin",
            codigo_point="LOCK002",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-lock-admin-001",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2025, 10, 1),
            month_end=date(2025, 10, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_PREVIOUS_CLOSURE,
            opening_reference_date=date(2025, 9, 30),
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=receta,
            inventario_inicial_teorico=Decimal("7"),
            inventario_final_teorico=Decimal("7"),
        )

        response = self.client.post(
            reverse("reportes:cierre_producto"),
            {"month": "2025-10", "action": "lock"},
            follow=True,
        )

        closure.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertTrue(closure.is_locked)
        self.assertEqual(closure.status, ProductoMonthClosure.STATUS_LOCKED)
        self.assertEqual(closure.metadata["lock_event"]["locked_by"], "admin_cierre")
        self.assertContains(response, "El cierre 2025-10 quedó bloqueado para proteger la conciliación mensual.")

    def test_cierre_producto_lock_rejects_catalog_issues(self):
        admin_user = User.objects.create_user(username="admin_cierre_issue", password="pass123")
        admin_group, _ = Group.objects.get_or_create(name="ADMIN")
        admin_user.groups.add(admin_group)
        self.client.force_login(admin_user)

        receta = Receta.objects.create(
            nombre="Pastel Lock Issue",
            codigo_point="LOCK003",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-lock-issue-001",
        )
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2025, 11, 1),
            month_end=date(2025, 11, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_PREVIOUS_CLOSURE,
            opening_reference_date=date(2025, 10, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=receta,
            inventario_inicial_teorico=Decimal("3"),
            inventario_final_teorico=Decimal("3"),
            has_catalog_issue=True,
            catalog_issue_note="Relacion derivada faltante",
        )

        response = self.client.post(
            reverse("reportes:cierre_producto"),
            {"month": "2025-11", "action": "lock"},
            follow=True,
        )

        closure.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertFalse(closure.is_locked)
        self.assertContains(response, "tiene incidencias de catalogo y no puede bloquearse")

    def _create_sucursal(self, codigo: str, nombre: str):
        from core.models import Sucursal

        return Sucursal.objects.create(codigo=codigo, nombre=nombre, activa=True)


class InventoryReportPresentationUnitTests(TestCase):
    def setUp(self):
        _clear_reportes_runtime_caches()
        self.user = User.objects.create_user(username="reportes_unidades_point", password="pass123")
        lectura, _ = Group.objects.get_or_create(name="LECTURA")
        self.user.groups.add(lectura)
        self.client.force_login(self.user)

        gramos = UnidadMedida.objects.create(
            codigo="g",
            nombre="Gramo",
            tipo=UnidadMedida.TIPO_MASA,
        )
        self.insumo = Insumo.objects.create(
            nombre="AZUCAR MASCABADO REPORTE",
            codigo_point="011-RPT",
            unidad_base=gramos,
            activo=True,
        )
        ExistenciaInsumo.objects.create(
            insumo=self.insumo,
            almacen="ALMACEN_1",
            stock_minimo=Decimal("30000"),
            stock_maximo=Decimal("100000"),
            punto_reorden=Decimal("60000"),
        )
        branch = PointBranch.objects.create(external_id="9", name="Almacen")
        job = PointSyncJob.objects.create(
            job_type=PointSyncJob.JOB_TYPE_INVENTORY,
            status=PointSyncJob.STATUS_SUCCESS,
            parameters={"canonical_insumo_inventory": True, "locations": ["ALMACEN", "CEDIS"]},
            result_summary={"locations": {"ALMACEN": {"rows": 1, "snapshots": 1}}},
        )
        PointInsumoInventorySnapshot.objects.create(
            branch=branch,
            insumo=self.insumo,
            point_code=self.insumo.codigo_point,
            point_name=self.insumo.nombre,
            point_quantity=Decimal("20"),
            point_unit="KG",
            quantity_base=Decimal("20000"),
            captured_at=timezone.now(),
            sync_job=job,
        )

    def test_faltantes_muestra_stock_y_reorden_en_kg(self):
        response = self.client.get(reverse("reportes:faltantes"), {"nivel": "all"})

        self.assertEqual(response.status_code, 200)
        row = next(item for item in response.context["rows"] if item.insumo_id == self.insumo.id)
        self.assertEqual(row.stock_actual_display, Decimal("20"))
        self.assertEqual(row.punto_reorden_display, Decimal("60"))
        self.assertEqual(row.sugerencia_compra_display, Decimal("40"))
        self.assertEqual(row.display_unit, "kg")
        self.assertContains(response, "20.000")
        self.assertContains(response, "60.000")
        self.assertContains(response, "40.000")
        self.assertNotContains(response, "20000.000")

    def test_faltantes_exporta_csv_y_xlsx_en_kg(self):
        csv_response = self.client.get(
            reverse("reportes:faltantes"),
            {"nivel": "all", "export": "csv"},
        )
        self.assertEqual(csv_response.status_code, 200)
        csv_text = csv_response.content.decode("utf-8")
        self.assertIn("AZUCAR MASCABADO REPORTE,kg,20.000,60.000,40.000", csv_text)
        self.assertNotIn("20000", csv_text)

        xlsx_response = self.client.get(
            reverse("reportes:faltantes"),
            {"nivel": "all", "export": "xlsx"},
        )
        self.assertEqual(xlsx_response.status_code, 200)
        workbook = load_workbook(BytesIO(xlsx_response.content), data_only=True)
        values = list(workbook["Faltantes"].iter_rows(values_only=True))
        data_row = next(row for row in values if row[0] == self.insumo.nombre)
        self.assertEqual(data_row[1:5], ("kg", 20, 60, 40))

    def test_auditoria_insumos_muestra_stock_y_politicas_en_kg(self):
        response = self.client.get(reverse("reportes:auditoria_insumos"))

        self.assertEqual(response.status_code, 200)
        stock = next(item for item in response.context["stock_rows"] if item.insumo_id == self.insumo.id)
        self.assertEqual(stock.stock_actual_display, Decimal("20"))
        self.assertEqual(stock.stock_minimo_display, Decimal("30"))
        self.assertEqual(stock.stock_maximo_display, Decimal("100"))
        self.assertEqual(stock.punto_reorden_display, Decimal("60"))
        self.assertEqual(stock.display_unit, "kg")
        self.assertContains(response, "20.000")
        self.assertContains(response, "30.000")
        self.assertContains(response, "100.000")
        self.assertContains(response, "60.000")
        self.assertNotContains(response, "20,000.000")
