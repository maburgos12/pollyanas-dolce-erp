from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from config import TIMEZONE
from hikconnect_client import CloudRecord


def load_export_file(path: str | Path) -> list[CloudRecord]:
    file_path = Path(path)
    if file_path.suffix.lower() == ".csv":
        return _load_csv(file_path)
    if file_path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _load_xlsx(file_path)
    raise ValueError("Formato no soportado. Usa CSV o XLSX exportado desde Hik-Connect Attendance.")


def _load_csv(path: Path) -> list[CloudRecord]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.DictReader(fh))
    return [_row_to_record(row) for row in rows if _row_to_record(row)]


def _load_xlsx(path: Path) -> list[CloudRecord]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    records = []
    for values in rows[1:]:
        row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
        record = _row_to_record(row)
        if record:
            records.append(record)
    return records


def _row_to_record(row: dict) -> CloudRecord | None:
    employee_no = str(row.get("ID") or row.get("id") or "").strip()
    date_s = str(row.get("Date") or row.get("date") or "").strip()
    time_s = str(row.get("Time") or row.get("time") or "").strip()
    if not employee_no or not date_s or not time_s:
        return None

    dt = _parse_export_datetime(date_s, time_s)
    first = str(row.get("First Name") or row.get("first_name") or "").strip()
    last = str(row.get("Last Name") or row.get("last_name") or "").strip()
    device_serial = str(row.get("Device Serial No.") or row.get("device_serial_no") or "").strip()
    record_guid = f"export:{employee_no}:{dt.isoformat()}:{device_serial}"
    return CloudRecord(
        record_guid=record_guid,
        employee_no=employee_no,
        name=" ".join(part for part in [first, last] if part),
        department=str(row.get("Department") or "").strip(),
        device_time=dt,
        device_name=str(row.get("Device Name") or "").strip(),
        device_serial_no=device_serial,
        raw=dict(row),
    )


def _parse_export_datetime(date_s: str, time_s: str) -> datetime:
    for fmt in ("%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(f"{date_s} {time_s}", fmt).replace(tzinfo=TIMEZONE)
        except ValueError:
            continue
    raise ValueError(f"No se pudo parsear fecha/hora exportada: {date_s} {time_s}")
