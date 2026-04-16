from __future__ import annotations

from datetime import date
import json
from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
from unittest.mock import patch

from django.core.management import call_command
from django.test import SimpleTestCase, TestCase
from openpyxl import load_workbook

from recetas.models import ProductoMonthClosure, ProductoMonthClosureLine, Receta
from pos_bridge.services.monthly_finished_goods_comparison_service import (
    CoverageSnapshot,
    MonthlyFinishedGoodsComparisonService,
)


class MonthlyFinishedGoodsComparisonServiceTests(TestCase):
    def setUp(self):
        self.service = MonthlyFinishedGoodsComparisonService()
        self.receta = Receta.objects.create(
            nombre="Pay de Queso Grande",
            codigo_point="PAY-Q-G",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-pay-queso-grande",
        )
        self.closure = ProductoMonthClosure.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
        )
        ProductoMonthClosureLine.objects.create(
            closure=self.closure,
            receta_padre=self.receta,
            inventario_inicial_teorico=10,
            produccion_mes=25,
            venta_directa_enteros=12,
            venta_derivada_equivalente=0,
            venta_total_equivalente=12,
            merma_total_equivalente=1,
            inventario_final_teorico=22,
        )

    def test_run_uses_point_fallback_when_month_has_no_rows(self):
        fake_sales_job = SimpleNamespace(id=101, status="SUCCESS", error_message="", result_summary={"rows": 44})
        fake_production_job = SimpleNamespace(id=102, status="SUCCESS", error_message="", result_summary={"rows": 12})
        fake_waste_job = SimpleNamespace(id=103, status="SUCCESS", error_message="", result_summary={"rows": 3})

        with TemporaryDirectory() as tmpdir, patch.object(
            self.service,
            "inspect_month_coverage",
            side_effect=[
                CoverageSnapshot(0, 0, 0, "", "", ""),
                CoverageSnapshot(44, 12, 3, "2026-03-31", "2026-03-31", "2026-03-31"),
            ],
        ), patch.object(
            self.service.official_sales_backfill_service,
            "run",
            return_value=fake_sales_job,
        ) as sales_mock, patch.object(
            self.service.movement_sync_service,
            "run_production_sync",
            return_value=fake_production_job,
        ) as production_mock, patch.object(
            self.service.movement_sync_service,
            "run_waste_sync",
            return_value=fake_waste_job,
        ) as waste_mock, patch.object(
            self.service.closure_service,
            "build",
            return_value=self.closure,
        ) as closure_mock:
            result = self.service.run(
                month="2026-03",
                output_dir=tmpdir,
                fallback_to_point=True,
            )
            self.assertTrue(Path(result["export_path"]).exists())
            wb = load_workbook(result["export_path"])
            self.assertIn("COMPARATIVA_DIRECTA", wb.sheetnames)
            self.assertIn("PRODUCCION_RAW", wb.sheetnames)
            self.assertIn("VENTAS_RAW", wb.sheetnames)
            self.assertIn("MERMAS_RAW", wb.sheetnames)

        sales_mock.assert_called_once()
        production_mock.assert_called_once()
        waste_mock.assert_called_once()
        closure_mock.assert_called_once()
        self.assertTrue(result["rebuild_applied"])
        self.assertEqual(len(result["sync_actions"]), 3)

    def test_run_uses_existing_closure_without_rebuild_when_coverage_exists(self):
        with TemporaryDirectory() as tmpdir, patch.object(
            self.service,
            "inspect_month_coverage",
            side_effect=[
                CoverageSnapshot(10, 5, 2, "2026-03-31", "2026-03-31", "2026-03-31"),
                CoverageSnapshot(10, 5, 2, "2026-03-31", "2026-03-31", "2026-03-31"),
            ],
        ), patch.object(
            self.service.closure_service,
            "build",
        ) as closure_mock:
            result = self.service.run(
                month="2026-03",
                output_dir=tmpdir,
                fallback_to_point=True,
                rebuild=False,
            )
            self.assertTrue(Path(result["export_path"]).exists())
            wb = load_workbook(result["export_path"])
            self.assertEqual(wb["RESUMEN_DIRECTO"]["B4"].value, "direct_raw_sources")

        closure_mock.assert_not_called()
        self.assertEqual(result["closure_id"], self.closure.id)
        self.assertEqual(result["sync_actions"], [])


class MonthlyFinishedGoodsComparisonCommandTests(SimpleTestCase):
    def test_command_prints_service_payload(self):
        payload = {
            "month": "2026-03",
            "closure_id": 17,
            "closure_status": "BUILT",
            "closure_locked": False,
            "coverage_before": {"sales_rows": 0},
            "coverage_after": {"sales_rows": 12},
            "sync_actions": [],
            "export_path": "/tmp/comparativa_producto_terminado_2026-03.xlsx",
        }

        with patch(
            "pos_bridge.management.commands.run_monthly_finished_goods_comparison.MonthlyFinishedGoodsComparisonService"
        ) as service_cls:
            service_cls.return_value.run.return_value = payload
            with TemporaryDirectory() as tmpdir:
                out_path = Path(tmpdir) / "stdout.txt"
                with out_path.open("w+", encoding="utf-8") as stream:
                    call_command(
                        "run_monthly_finished_goods_comparison",
                        month="2026-03",
                        output_dir=tmpdir,
                        stdout=stream,
                    )
                    stream.seek(0)
                    rendered = stream.read()

        self.assertEqual(json.loads(rendered), payload)
