from __future__ import annotations

import json
from dataclasses import replace
from datetime import date, datetime
from decimal import Decimal
from io import StringIO
from tempfile import NamedTemporaryFile
from types import MappingProxyType
from unittest.mock import patch

from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase
from django.db import connection
from django.test.utils import CaptureQueriesContext, override_settings
from django.utils import timezone
from openpyxl import Workbook

from core.models import Sucursal
from pos_bridge.models import PointBranch, PointDailySale, PointInventorySnapshot, PointProduct, PointSyncJob, PointWasteLine
from pos_bridge.services.product_month_closure_service import ProductMonthClosureError, ProductMonthClosureService
from pos_bridge.services.monthly_product_balance_service import (
    MonthlyPointBalance,
    MonthlyPointBalanceRow,
    MonthlyPointProductBalanceService,
)
from recetas.models import (
    ProductoMonthClosure,
    ProductoMonthClosureLine,
    Receta,
    RecetaEquivalencia,
    RecetaPresentacionDerivada,
    VentaHistorica,
)
from reportes.models import FactProduccionDiaria
from pos_bridge.models.movements import PointProductionLine


@override_settings(PRODUCT_MONTH_CLOSURE_SALES_SOURCE_MODE="BRIDGE_HISTORY")
class ProductMonthClosureServiceTests(TestCase):
    def setUp(self):
        self.service = ProductMonthClosureService()
        self.sucursal, _ = Sucursal.objects.update_or_create(
            codigo="CEDIS",
            defaults={"nombre": "CEDIS", "activa": True},
        )
        self.point_branch, _ = PointBranch.objects.get_or_create(
            external_id="CEDIS",
            defaults={"name": "CEDIS", "erp_branch": self.sucursal},
        )
        if self.point_branch.erp_branch_id != self.sucursal.id:
            self.point_branch.erp_branch = self.sucursal
            self.point_branch.save(update_fields=["erp_branch"])
        self.sync_job = PointSyncJob.objects.create(
            job_type=PointSyncJob.JOB_TYPE_INVENTORY,
            status=PointSyncJob.STATUS_SUCCESS,
        )
        self.parent = Receta.objects.create(
            nombre="Pastel de Snickers Mediano",
            codigo_point="SNK-M",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-parent-snk-mediano",
        )
        self.derived = Receta.objects.create(
            nombre="Pastel de Snickers Rebanada",
            codigo_point="SNK-R",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-derived-snk-rebanada",
        )
        self.derived_relation = RecetaPresentacionDerivada.objects.create(
            receta_padre=self.parent,
            receta_derivada=self.derived,
            codigo_point_derivado="SNK-R",
            nombre_derivado=self.derived.nombre,
            unidades_por_padre=Decimal("10"),
        )

    def _canonical_balance(self, rows, *, issues=(), selected_dates=(date(2025, 9, 30),)):
        return MonthlyPointBalance(
            month_start=date(2025, 9, 1),
            month_end=date(2025, 9, 30),
            rows=MappingProxyType(rows),
            issues=tuple(issues),
            sources=MappingProxyType(
                {
                    "opening_snapshot": MappingProxyType(
                        {"authoritative": True, "selected_dates": tuple(selected_dates)}
                    ),
                    "closing_snapshot": MappingProxyType(
                        {"authoritative": True, "selected_dates": tuple(selected_dates)}
                    ),
                    "sales": MappingProxyType(
                        {"configured_source_mode": "BRIDGE_HISTORY", "selected_source": "bridge_history"}
                    ),
                }
            ),
            effective_snapshot_dates=MappingProxyType({"opening": None, "closing": None}),
            source_counts=MappingProxyType({"opening_snapshot_rows": 2, "closing_snapshot_rows": 2}),
        )

    def _lockable_fingerprint_balance(self, *, sales_job_id=101):
        source = lambda **extra: MappingProxyType(
            {
                "source_present": True,
                "authoritative": True,
                **extra,
            }
        )
        return MonthlyPointBalance(
            month_start=date(2025, 9, 1),
            month_end=date(2025, 9, 30),
            rows=MappingProxyType(
                {
                    self.parent.id: MonthlyPointBalanceRow(
                        receta_id=self.parent.id,
                        opening_point=Decimal("10"),
                        production=Decimal("2"),
                        sales=Decimal("1"),
                        waste=Decimal("0"),
                        conversion_in=Decimal("0"),
                        conversion_out=Decimal("0"),
                        calculated_closing=Decimal("11"),
                        closing_point=Decimal("11"),
                        closing_point_cedis=Decimal("6"),
                        closing_point_sucursales=Decimal("5"),
                        difference_point=Decimal("0"),
                        status="COINCIDE",
                        source_counts=MappingProxyType(
                            {
                                "opening_snapshot_rows": 1,
                                "closing_snapshot_rows": 1,
                                "production_rows": 1,
                                "sales_rows": 1,
                            }
                        ),
                    )
                }
            ),
            sources=MappingProxyType(
                {
                    "opening_snapshot": source(
                        selected_dates=(date(2025, 8, 31),),
                        effective_date=date(2025, 8, 31),
                        selected_sync_job_ids=(91,),
                        snapshot_rows=1,
                    ),
                    "closing_snapshot": source(
                        selected_dates=(date(2025, 9, 30),),
                        effective_date=date(2025, 9, 30),
                        selected_sync_job_ids=(92,),
                        snapshot_rows=1,
                    ),
                    "sales": source(
                        selected_source="official_point_daily_sales",
                        job_id=sales_job_id,
                        job_status=PointSyncJob.STATUS_SUCCESS,
                        official_daily_row_count=1,
                    ),
                    "production": source(selected_sync_job_ids=(93,), row_count=1),
                    "waste": source(selected_sync_job_ids=(94,), row_count=0),
                    "conversions": source(selected_sync_job_ids=(95,), row_count=0),
                }
            ),
            effective_snapshot_dates=MappingProxyType(
                {"opening": date(2025, 8, 31), "closing": date(2025, 9, 30)}
            ),
            source_counts=MappingProxyType(
                {
                    "opening_snapshot_rows": 1,
                    "closing_snapshot_rows": 1,
                    "production_rows": 1,
                    "sales_rows": 1,
                    "waste_rows": 0,
                    "conversion_rows": 0,
                }
            ),
        )

    def test_projection_persists_each_source_authority_even_when_source_is_present(self):
        class CanonicalBalance:
            def __init__(self, balance):
                self.balance = balance

            def build(self, month, **kwargs):
                return self.balance

        balance = self._lockable_fingerprint_balance()
        sources = dict(balance.sources)
        sources["sales"] = MappingProxyType(
            {**dict(sources["sales"]), "source_present": True, "authoritative": False}
        )
        sources["closing_snapshot"] = MappingProxyType(
            {**dict(sources["closing_snapshot"]), "source_present": True, "authoritative": False}
        )
        balance = replace(balance, sources=MappingProxyType(sources), issues=("MONTH_SOURCE_INCOMPLETE",))

        preview = ProductMonthClosureService(balance_service=CanonicalBalance(balance)).preview(month="2025-09")

        metadata = preview["line_rows"][0]["metadata"]
        self.assertTrue(metadata["opening_source_authoritative"])
        self.assertFalse(metadata["sales_source_authoritative"])
        self.assertTrue(metadata["production_source_authoritative"])
        self.assertTrue(metadata["waste_source_authoritative"])
        self.assertTrue(metadata["conversion_source_authoritative"])
        self.assertFalse(metadata["closing_source_authoritative"])

    def test_preview_projects_canonical_rows_to_parent_and_preserves_json_metadata(self):
        class CanonicalBalance:
            def __init__(self, balance):
                self.balance = balance

            def build(self, month, **kwargs):
                return self.balance

        RecetaEquivalencia.objects.create(
            receta_porcion=self.derived,
            receta_padre=self.parent,
            factor_conversion=Decimal("8"),
            tipo_relacion=RecetaEquivalencia.TIPO_CONVERSION,
            activo=True,
        )

        parent_row = MonthlyPointBalanceRow(
            receta_id=self.parent.id,
            opening_point=Decimal("10"),
            production=Decimal("3"),
            sales=Decimal("1"),
            waste=Decimal("0"),
            conversion_out=Decimal("2"),
            calculated_closing=Decimal("10"),
            closing_point=Decimal("8"),
            closing_point_cedis=Decimal("5"),
            closing_point_sucursales=Decimal("3"),
            difference_point=Decimal("0"),
            status="COINCIDE",
            conversion_origin="POINT",
            conversion_origins=("POINT",),
            source_counts=MappingProxyType({"opening_snapshot_rows": 1, "conversion_out_rows": 1}),
        )
        slice_row = MonthlyPointBalanceRow(
            receta_id=self.derived.id,
            opening_point=Decimal("0"),
            sales=Decimal("16"),
            conversion_in=Decimal("16"),
            calculated_closing=Decimal("0"),
            closing_point=Decimal("16"),
            closing_point_cedis=Decimal("8"),
            closing_point_sucursales=Decimal("8"),
            difference_point=Decimal("0"),
            status="COINCIDE",
            conversion_origin="EQUIVALENCIA_CONFIGURADA",
            conversion_origins=("EQUIVALENCIA_CONFIGURADA",),
            source_counts=MappingProxyType({"sales_rows": 1, "conversion_in_rows": 1}),
        )
        service = ProductMonthClosureService(
            balance_service=CanonicalBalance(self._canonical_balance({self.parent.id: parent_row, self.derived.id: slice_row}))
        )

        preview = service.preview(month="2025-09")

        row = preview["line_rows"][0]
        self.assertEqual(row["receta"], self.parent)
        self.assertEqual(row["inventario_inicial_teorico"], Decimal("10"))
        self.assertEqual(row["produccion_mes"], Decimal("3"))
        self.assertEqual(row["venta_directa_enteros"], Decimal("1"))
        self.assertEqual(row["venta_derivada_equivalente"], Decimal("2"))
        self.assertEqual(row["inventario_final_teorico"], Decimal("10"))
        self.assertEqual(row["diferencia_teorico_vs_point"], Decimal("0"))
        self.assertEqual(row["metadata"]["point_conversion_in"], "2")
        self.assertEqual(row["metadata"]["point_conversion_out"], "2")
        self.assertEqual(row["metadata"]["conversion_origin"], "MIXED")
        self.assertEqual(
            row["metadata"]["conversion_origins"],
            ["EQUIVALENCIA_CONFIGURADA", "POINT"],
        )
        self.assertEqual(row["metadata"]["projection_sources"], ["DIRECTA", "EQUIVALENCIA"])
        self.assertEqual(row["metadata"]["balance_contract"], "POINT_PRODUCT_BALANCE_V1")
        self.assertTrue(row["metadata"]["point_final_scopes_available"])
        self.assertEqual(row["inventario_final_point_cedis"], Decimal("6"))
        self.assertEqual(row["inventario_final_point_sucursales"], Decimal("4"))
        self.assertEqual(row["inventario_final_point_total"], Decimal("10"))

        closure = service.build(month="2025-09")
        persisted = closure.lines.get(receta_padre=self.parent)
        self.assertEqual(persisted.metadata["source_counts"]["conversion_in_rows"], 1)
        self.assertEqual(
            persisted.inventario_final_point_total,
            persisted.inventario_final_point_cedis + persisted.inventario_final_point_sucursales,
        )
        self.assertEqual(preview["totals"]["closing_cedis"], Decimal("6"))
        self.assertEqual(preview["totals"]["closing_sucursales"], Decimal("4"))
        self.assertEqual(preview["totals"]["closing_total"], Decimal("10"))

    def test_canonical_closure_preserves_unresolved_origin_without_inference(self):
        balance = self._canonical_balance(
            {
                self.parent.id: MonthlyPointBalanceRow(
                    receta_id=self.parent.id,
                    opening_point=Decimal("1"),
                    sales=Decimal("0"),
                    conversion_in=Decimal("1"),
                    calculated_closing=Decimal("2"),
                    closing_point=Decimal("2"),
                    closing_point_cedis=Decimal("2"),
                    closing_point_sucursales=Decimal("0"),
                    difference_point=Decimal("0"),
                    conversion_origin="UNRESOLVED",
                    conversion_origins=("UNRESOLVED",),
                    issues=("CONVERSION_ORIGIN_UNRESOLVED",),
                )
            },
            issues=("CONVERSION_ORIGIN_UNRESOLVED",),
        )

        rows = self.service._project_canonical_balance(balance=balance)

        self.assertEqual(rows[0]["metadata"]["conversion_origin"], "UNRESOLVED")
        self.assertEqual(rows[0]["metadata"]["conversion_origins"], ["UNRESOLVED"])

    def test_projection_trace_only_includes_rows_with_conversion_movements(self):
        balance = self._canonical_balance(
            {
                self.parent.id: MonthlyPointBalanceRow(
                    receta_id=self.parent.id,
                    conversion_out=Decimal("1"),
                    conversion_origin="POINT",
                    conversion_origins=("POINT",),
                ),
                self.derived.id: MonthlyPointBalanceRow(
                    receta_id=self.derived.id,
                    sales=Decimal("10"),
                ),
            }
        )

        rows = self.service._project_canonical_balance(balance=balance)

        self.assertEqual(rows[0]["metadata"]["projection_sources"], ["DIRECTA"])

    def test_costing_equivalence_never_projects_or_scales_operational_balance(self):
        costing_parent = Receta.objects.create(
            nombre="Base de costeo",
            codigo_point="COST-P",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-cost-parent",
        )
        costing_child = Receta.objects.create(
            nombre="Producto con complemento de costo",
            codigo_point="COST-C",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-cost-child",
        )
        RecetaEquivalencia.objects.create(
            receta_porcion=costing_child,
            receta_padre=costing_parent,
            factor_conversion=Decimal("8"),
            tipo_relacion=RecetaEquivalencia.TIPO_COSTEO,
            activo=True,
        )
        balance = self._canonical_balance(
            {
                costing_child.id: MonthlyPointBalanceRow(
                    receta_id=costing_child.id,
                    opening_point=Decimal("3"),
                    sales=Decimal("8"),
                    calculated_closing=Decimal("1"),
                    closing_point=Decimal("1"),
                    difference_point=Decimal("0"),
                    status="COINCIDE",
                )
            }
        )

        rows = self.service._project_canonical_balance(balance=balance)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["receta"], costing_child)
        self.assertEqual(rows[0]["inventario_inicial_teorico"], Decimal("3"))
        self.assertEqual(rows[0]["venta_directa_enteros"], Decimal("8"))
        self.assertEqual(rows[0]["venta_derivada_equivalente"], Decimal("0"))
        self.assertEqual(rows[0]["metadata"]["projection_sources"], [])

    def test_new_canonical_audit_detail_names_point_balance_not_physical_inventory(self):
        missing_status, missing_detail = self.service._canonical_audit_status(
            issues=set(),
            closing_missing=True,
            point_difference=None,
            waste_total=Decimal("0"),
        )
        greater_status, greater_detail = self.service._canonical_audit_status(
            issues=set(),
            closing_missing=False,
            point_difference=Decimal("2"),
            waste_total=Decimal("0"),
        )
        lower_status, lower_detail = self.service._canonical_audit_status(
            issues=set(),
            closing_missing=False,
            point_difference=Decimal("-2"),
            waste_total=Decimal("0"),
        )

        self.assertEqual(missing_status, ProductoMonthClosureLine.AUDIT_STATUS_SIN_INVENTARIO_FISICO)
        self.assertIn("inventario final Point", missing_detail)
        for detail in (missing_detail, greater_detail, lower_detail):
            self.assertNotIn("físic", detail.lower())
            self.assertNotIn("fisic", detail.lower())

    def test_preview_sales_authority_alone_controls_lock_readiness_and_compacts_metadata(self):
        class CanonicalBalance:
            def __init__(self, balance): self.balance = balance
            def build(self, month, **kwargs): return self.balance

        row = MonthlyPointBalanceRow(
            receta_id=self.parent.id, opening_point=Decimal("1"), calculated_closing=Decimal("1"),
            closing_point=Decimal("1"), closing_point_cedis=Decimal("1"), closing_point_sucursales=Decimal("0"),
            difference_point=Decimal("0"), status="COINCIDE", source_counts=MappingProxyType({"closing_snapshot_rows": 1}),
        )
        balance = self._canonical_balance({self.parent.id: row}, issues=("MONTH_SOURCE_INCOMPLETE",))
        sources = dict(balance.sources)
        sources["sales"] = MappingProxyType({
            "authoritative": False, "selected_source": "bridge_history", "selected_row_job_ids": tuple(range(30)),
            "recipe_scope_totals": {str(index): index for index in range(30)},
        })
        balance = MonthlyPointBalance(
            month_start=balance.month_start, month_end=balance.month_end, rows=balance.rows, issues=balance.issues,
            sources=MappingProxyType(sources), effective_snapshot_dates=balance.effective_snapshot_dates,
            source_counts=balance.source_counts,
        )
        preview = ProductMonthClosureService(balance_service=CanonicalBalance(balance)).preview(month="2025-09")
        self.assertFalse(preview["metadata"]["validation"]["lock_ready"])
        self.assertIn("MONTH_SOURCE_INCOMPLETE", preview["metadata"]["validation"]["blocking_issues"])
        compact = preview["metadata"]["sales_meta"]
        self.assertEqual(compact["selected_row_job_ids"]["count"], 30)
        self.assertEqual(compact["recipe_scope_totals"]["count"], 30)

    def test_preview_blocks_each_required_source_marked_absent_even_if_authoritative_flag_is_stale(self):
        class CanonicalBalance:
            def __init__(self, balance): self.balance = balance
            def build(self, month, **kwargs): return self.balance

        for absent_family in ("sales", "production", "waste", "conversions"):
            with self.subTest(absent_family=absent_family):
                sources = {
                    family: MappingProxyType({"authoritative": True, "source_present": family != absent_family})
                    for family in (
                        "opening_snapshot",
                        "closing_snapshot",
                        "sales",
                        "production",
                        "waste",
                        "conversions",
                    )
                }
                balance = MonthlyPointBalance(
                    month_start=date(2025, 9, 1),
                    month_end=date(2025, 9, 30),
                    rows=MappingProxyType({
                        self.parent.id: MonthlyPointBalanceRow(
                            receta_id=self.parent.id,
                            opening_point=Decimal("10"),
                            calculated_closing=Decimal("10"),
                            closing_point=Decimal("10"),
                            difference_point=Decimal("0"),
                            status="COINCIDE",
                        )
                    }),
                    sources=MappingProxyType(sources),
                    effective_snapshot_dates=MappingProxyType(
                        {"opening": date(2025, 8, 31), "closing": date(2025, 9, 30)}
                    ),
                )
                self.service.balance_service = CanonicalBalance(balance)

                preview = self.service.preview(month="2025-09")

                self.assertIn("MONTH_SOURCE_INCOMPLETE", preview["metadata"]["validation"]["blocking_issues"])
                self.assertFalse(preview["metadata"]["validation"]["lock_ready"])

    def test_preview_propagates_sales_authority_matrix_to_lock_readiness(self):
        class CanonicalBalance:
            def __init__(self, balance):
                self.balance = balance

            def build(self, month, **kwargs):
                return self.balance

        row = MonthlyPointBalanceRow(
            receta_id=self.parent.id,
            opening_point=Decimal("10"),
            production=Decimal("2"),
            sales=Decimal("3"),
            waste=Decimal("1"),
            calculated_closing=Decimal("8"),
            closing_point=Decimal("8"),
            closing_point_cedis=Decimal("5"),
            closing_point_sucursales=Decimal("3"),
            difference_point=Decimal("0"),
            status="COINCIDE",
            source_counts=MappingProxyType(
                {
                    "opening_snapshot_rows": 1,
                    "closing_snapshot_rows": 1,
                    "sales_rows": 1,
                    "production_rows": 1,
                    "waste_rows": 1,
                }
            ),
        )
        base_sources = {
            "opening_snapshot": MappingProxyType({"authoritative": True, "selected_dates": (date(2025, 8, 31),)}),
            "closing_snapshot": MappingProxyType({"authoritative": True, "selected_dates": (date(2025, 9, 30),)}),
        }
        cases = (
            ("bridge", False, "SALES_SOURCE_REQUIRES_REVIEW", {"selected_source": "bridge_history"}),
            ("missing_job", False, "SALES_SYNC_JOB_MISSING", {"selected_source": "official_point_daily_sales"}),
            ("partial_job", False, "SALES_SYNC_JOB_PARTIAL", {"selected_source": "official_point_daily_sales", "job_status": "PARTIAL"}),
            ("failed_job", False, "SALES_SYNC_JOB_FAILED", {"selected_source": "official_point_daily_sales", "job_status": "FAILED"}),
            ("restricted_job", False, "SALES_SYNC_JOB_RESTRICTED", {"selected_source": "official_point_daily_sales"}),
            ("coverage_unproven", False, "SALES_SYNC_COVERAGE_UNPROVEN", {"selected_source": "official_point_daily_sales"}),
            ("divergent_bridge", False, "SALES_SOURCE_MIXED", {"selected_source": "official_point_daily_sales", "materialized_bridge_reconciled": False}),
            ("official_full", True, None, {"selected_source": "official_point_daily_sales", "job_id": 901, "job_status": "SUCCESS"}),
            ("reconciled_duplicate", True, None, {"selected_source": "official_point_daily_sales", "job_id": 902, "job_status": "SUCCESS", "materialized_bridge_reconciled": True}),
        )

        for label, authoritative, expected_issue, extra_sales in cases:
            with self.subTest(case=label):
                sales = {
                    "configured_source_mode": "AUTO",
                    "selected_source": "official_point_daily_sales",
                    "authoritative": authoritative,
                    "selected_dates": (date(2025, 9, 1), date(2025, 9, 30)),
                    "coverage_expected_branch_days": 30,
                    "coverage_logged_branch_days": 30,
                    **extra_sales,
                }
                balance = MonthlyPointBalance(
                    month_start=date(2025, 9, 1),
                    month_end=date(2025, 9, 30),
                    rows=MappingProxyType({self.parent.id: row}),
                    issues=() if expected_issue is None else (expected_issue,),
                    sources=MappingProxyType({**base_sources, "sales": MappingProxyType(sales)}),
                    effective_snapshot_dates=MappingProxyType({"opening": date(2025, 8, 31), "closing": date(2025, 9, 30)}),
                    source_counts=MappingProxyType({"opening_snapshot_rows": 1, "closing_snapshot_rows": 1, "sales_rows": 1}),
                )

                preview = ProductMonthClosureService(balance_service=CanonicalBalance(balance)).preview(month="2025-09")
                validation = preview["metadata"]["validation"]

                self.assertEqual(validation["lock_ready"], expected_issue is None)
                if expected_issue is None:
                    self.assertEqual(validation["blocking_issues"], [])
                else:
                    self.assertIn(expected_issue, validation["blocking_issues"])

    def test_preview_copies_sales_validation_compatibility_fields_from_canonical_metadata(self):
        class CanonicalBalance:
            def __init__(self, balance):
                self.balance = balance

            def build(self, month, **kwargs):
                return self.balance

        row = MonthlyPointBalanceRow(
            receta_id=self.parent.id,
            opening_point=Decimal("10"),
            calculated_closing=Decimal("10"),
            closing_point=Decimal("10"),
            closing_point_cedis=Decimal("10"),
            closing_point_sucursales=Decimal("0"),
            difference_point=Decimal("0"),
            status="COINCIDE",
            source_counts=MappingProxyType({"opening_snapshot_rows": 1, "closing_snapshot_rows": 1, "sales_rows": 12}),
        )
        balance = MonthlyPointBalance(
            month_start=date(2025, 9, 1),
            month_end=date(2025, 9, 30),
            rows=MappingProxyType({self.parent.id: row}),
            sources=MappingProxyType(
                {
                    "opening_snapshot": MappingProxyType({"authoritative": True, "selected_dates": (date(2025, 8, 31),)}),
                    "closing_snapshot": MappingProxyType({"authoritative": True, "selected_dates": (date(2025, 9, 30),)}),
                    "sales": MappingProxyType(
                        {
                            "authoritative": True,
                            "selected_source": "official_point_daily_sales",
                            "job_id": 445,
                            "job_status": "SUCCESS",
                            "official_daily_row_count": 12,
                            "legacy_daily_row_count": 3,
                            "legacy_bridge_row_count": 4,
                        }
                    ),
                }
            ),
            effective_snapshot_dates=MappingProxyType({"opening": date(2025, 8, 31), "closing": date(2025, 9, 30)}),
            source_counts=MappingProxyType({"opening_snapshot_rows": 1, "closing_snapshot_rows": 1, "sales_rows": 999}),
        )

        validation = ProductMonthClosureService(balance_service=CanonicalBalance(balance)).preview(month="2025-09")["metadata"]["validation"]

        self.assertEqual(validation["sales_job_id"], 445)
        self.assertEqual(validation["sales_job_status"], "SUCCESS")
        self.assertEqual(validation["sales_official_rows"], 12)
        self.assertEqual(validation["sales_legacy_rows"], 7)

    def test_preview_preserves_complete_compact_movement_authority_metadata(self):
        class CanonicalBalance:
            def __init__(self, balance):
                self.balance = balance

            def build(self, month, **kwargs):
                return self.balance

        row = MonthlyPointBalanceRow(
            receta_id=self.parent.id,
            opening_point=Decimal("10"),
            calculated_closing=None,
            closing_point=Decimal("10"),
            difference_point=None,
            status="REVISAR_FUENTE",
            issues=("CALCULATED_CLOSING_MISSING",),
        )
        movement_sources = {
            "production": {
                "source": "PointProductionLine",
                "authoritative": False,
                "source_present": True,
                "job_status": "PARTIAL",
                "selected_sync_job_ids": tuple(range(20, 40)),
                "authority_issues": (
                    "PRODUCTION_SYNC_JOB_PARTIAL",
                    "PRODUCTION_SYNC_RANGE_INCOMPLETE",
                    "PRODUCTION_SYNC_JOB_RESTRICTED",
                    "PRODUCTION_SYNC_CONTRACT_INCOMPLETE",
                    "PRODUCTION_SYNC_COUNT_MISMATCH",
                    "PRODUCTION_SYNC_JOB_MIXED",
                ),
            },
            "waste": {
                "source": "PointWasteLine",
                "authoritative": False,
                "source_present": False,
                "authority_issues": ("WASTE_SYNC_JOB_MISSING",),
            },
            "conversions": {
                "source": "PointConversionLine",
                "authoritative": False,
                "source_present": True,
                "authority_issues": ("CONVERSION_SYNC_BRANCH_COVERAGE_INCOMPLETE",),
            },
        }
        sources = dict(self._canonical_balance({self.parent.id: row}).sources)
        sources.update({key: MappingProxyType(value) for key, value in movement_sources.items()})
        balance = MonthlyPointBalance(
            month_start=date(2025, 9, 1),
            month_end=date(2025, 9, 30),
            rows=MappingProxyType({self.parent.id: row}),
            issues=("MONTH_SOURCE_INCOMPLETE",),
            sources=MappingProxyType(sources),
            effective_snapshot_dates=MappingProxyType(
                {"opening": date(2025, 8, 31), "closing": date(2025, 9, 30)}
            ),
            source_counts=MappingProxyType({}),
        )

        preview = ProductMonthClosureService(balance_service=CanonicalBalance(balance)).preview(month="2025-09")

        metadata = preview["metadata"]
        self.assertEqual(
            metadata["production_meta"]["authority_issues"],
            list(movement_sources["production"]["authority_issues"]),
        )
        self.assertEqual(metadata["waste_meta"]["authority_issues"], ["WASTE_SYNC_JOB_MISSING"])
        self.assertEqual(
            metadata["conversion_meta"]["authority_issues"],
            ["CONVERSION_SYNC_BRANCH_COVERAGE_INCOMPLETE"],
        )
        self.assertEqual(metadata["production_meta"]["selected_sync_job_ids"]["count"], 20)
        self.assertEqual(metadata["production_meta"]["job_status"], "PARTIAL")

    def test_preview_recursively_compacts_large_source_metadata_with_stable_full_hashes(self):
        class CanonicalBalance:
            def __init__(self, balance):
                self.balance = balance

            def build(self, month, **kwargs):
                return self.balance

        row = MonthlyPointBalanceRow(
            receta_id=self.parent.id,
            opening_point=Decimal("10"),
            calculated_closing=Decimal("10"),
            closing_point=Decimal("10"),
            closing_point_cedis=Decimal("10"),
            closing_point_sucursales=Decimal("0"),
            difference_point=Decimal("0"),
            status="COINCIDE",
            source_counts=MappingProxyType({"opening_snapshot_rows": 1, "closing_snapshot_rows": 1, "sales_rows": 1}),
        )
        source = MappingProxyType(
            {
                "configured_source_mode": "AUTO",
                "selected_source": "official_point_daily_sales",
                "authoritative": True,
                "job_id": 777,
                "selected_dates": tuple(date(2025, 9, day) for day in range(1, 11)),
                "warnings": tuple(f"advertencia sintetica {index}" for index in range(40)),
                "coverage_expected_branch_days": 270,
                "coverage_logged_branch_days": 270,
                "selected_row_job_ids": tuple(range(100, 140)),
                "coverage_no_aplica_branch_days": tuple(f"CEDIS:{day:02d}" for day in range(40)),
                "coverage_missing_branch_days": tuple(f"SUC:{day:02d}" for day in range(40)),
                "rejected_provenance": tuple({"job_id": job_id, "reason": "synthetic"} for job_id in range(40)),
                "recipe_scope_totals": {str(index): {"cedis": Decimal(index), "sucursales": Decimal(index + 1)} for index in range(40)},
                "nested": {"mixed": [{"positions": list(range(40)), "labels": [f"r-{index}" for index in range(40)]}]},
            }
        )
        sources = MappingProxyType(
            {
                "opening_snapshot": MappingProxyType({"authoritative": True, "selected_dates": (date(2025, 8, 31),)}),
                "closing_snapshot": MappingProxyType({"authoritative": True, "selected_dates": (date(2025, 9, 30),)}),
                "sales": source,
            }
        )
        balance = MonthlyPointBalance(
            month_start=date(2025, 9, 1),
            month_end=date(2025, 9, 30),
            rows=MappingProxyType({self.parent.id: row}),
            sources=sources,
            effective_snapshot_dates=MappingProxyType({"opening": date(2025, 8, 31), "closing": date(2025, 9, 30)}),
            source_counts=MappingProxyType({"opening_snapshot_rows": 1, "closing_snapshot_rows": 1, "sales_rows": 1}),
        )
        service = ProductMonthClosureService(balance_service=CanonicalBalance(balance))

        preview = service.preview(month="2025-09")
        closure = service.build(month="2025-09")
        metadata = closure.metadata
        sales_meta = metadata["sales_meta"]
        serialized = json.dumps(metadata, sort_keys=True).encode("utf-8")

        for key, original in {
            "selected_row_job_ids": source["selected_row_job_ids"],
            "coverage_no_aplica_branch_days": source["coverage_no_aplica_branch_days"],
            "coverage_missing_branch_days": source["coverage_missing_branch_days"],
            "rejected_provenance": source["rejected_provenance"],
            "recipe_scope_totals": source["recipe_scope_totals"],
        }.items():
            with self.subTest(key=key):
                compact = sales_meta[key]
                self.assertIsInstance(compact, dict)
                self.assertEqual(compact["count"], len(original))
                self.assertRegex(compact["hash"], r"^[0-9a-f]{64}$")
                self.assertLessEqual(len(compact["sample"]), service.METADATA_COMPACTION_SAMPLE_LIMIT)
                self.assertNotEqual(compact, original)
                self.assertNotEqual(compact["sample"], ProductMonthClosureService._json_compatible(original))

        nested_positions = sales_meta["nested"]["mixed"][0]["positions"]
        self.assertEqual(nested_positions["count"], 40)
        self.assertRegex(nested_positions["hash"], r"^[0-9a-f]{64}$")
        self.assertLessEqual(len(nested_positions["sample"]), service.METADATA_COMPACTION_SAMPLE_LIMIT)
        self.assertEqual(sales_meta["configured_source_mode"], "AUTO")
        self.assertEqual(sales_meta["selected_source"], "official_point_daily_sales")
        self.assertTrue(sales_meta["authoritative"])
        self.assertEqual(sales_meta["job_id"], 777)
        self.assertEqual(sales_meta["selected_dates"], [f"2025-09-{day:02d}" for day in range(1, 11)])
        self.assertEqual(sales_meta["warnings"], [f"advertencia sintetica {index}" for index in range(40)])
        self.assertIsInstance(sales_meta["selected_dates"], list)
        self.assertIsInstance(sales_meta["warnings"], list)
        self.assertEqual(sales_meta["coverage_expected_branch_days"], 270)
        self.assertEqual(sales_meta["coverage_logged_branch_days"], 270)
        self.assertIn("opening_meta", metadata)
        self.assertIn("sales_meta", metadata)
        self.assertIn("closing_inventory_meta", metadata)
        self.assertLess(len(serialized), 20_000)
        self.assertEqual(preview["metadata"]["sales_meta"], sales_meta)

    def test_project_canonical_balance_uses_bounded_queries_and_equivalence_precedence(self):
        def balance_for(recipes):
            return MonthlyPointBalance(
                month_start=date(2025, 9, 1),
                month_end=date(2025, 9, 30),
                rows=MappingProxyType(
                    {
                        receta.id: MonthlyPointBalanceRow(
                            receta_id=receta.id,
                            opening_point=Decimal("20"),
                            sales=Decimal("20") if receta != self.parent else Decimal("3"),
                            calculated_closing=Decimal("20") if receta != self.parent else Decimal("17"),
                            closing_point=Decimal("20") if receta != self.parent else Decimal("17"),
                            closing_point_cedis=Decimal("20") if receta != self.parent else Decimal("17"),
                            closing_point_sucursales=Decimal("0"),
                            difference_point=Decimal("0"),
                            status="COINCIDE",
                            source_counts=MappingProxyType({"opening_snapshot_rows": 1, "closing_snapshot_rows": 1, "sales_rows": 1}),
                        )
                        for receta in recipes
                    }
                ),
            )

        raw_recipes = []
        for index in range(20):
            raw = Receta.objects.create(
                nombre=f"Producto proyeccion {index}",
                codigo_point=f"PROY-{index}",
                tipo=Receta.TIPO_PRODUCTO_FINAL,
                hash_contenido=f"projection-{index}",
            )
            raw_recipes.append(raw)
            if index % 2 == 0:
                RecetaEquivalencia.objects.create(
                    receta_porcion=raw,
                    receta_padre=self.parent,
                    factor_conversion=Decimal("2"),
                    tipo_relacion=RecetaEquivalencia.TIPO_CONVERSION,
                    activo=True,
                )
            else:
                RecetaPresentacionDerivada.objects.create(
                    receta_padre=self.parent,
                    receta_derivada=raw,
                    codigo_point_derivado=f"PROY-DER-{index}",
                    nombre_derivado=raw.nombre,
                    unidades_por_padre=Decimal("10"),
                    activo=True,
                )
        hybrid = raw_recipes[0]
        RecetaPresentacionDerivada.objects.create(
            receta_padre=self.parent,
            receta_derivada=hybrid,
            codigo_point_derivado="PROY-HYBRID",
            nombre_derivado=hybrid.nombre,
            unidades_por_padre=Decimal("10"),
            activo=True,
        )
        service = ProductMonthClosureService()

        with CaptureQueriesContext(connection) as small_queries:
            small_rows = service._project_canonical_balance(balance=balance_for([self.parent]))
        with CaptureQueriesContext(connection) as many_queries:
            many_rows = service._project_canonical_balance(balance=balance_for([self.parent, *raw_recipes]))

        self.assertEqual(len(small_queries), 3)
        self.assertEqual(len(many_queries), 3)
        self.assertLessEqual(len(many_queries), len(small_queries) + 1)
        self.assertEqual(len(many_rows), 1)
        self.assertEqual(many_rows[0]["receta"], self.parent)
        self.assertEqual(many_rows[0]["venta_directa_enteros"], Decimal("3"))
        self.assertEqual(many_rows[0]["venta_derivada_equivalente"], Decimal("120"))
        self.assertEqual(small_rows[0]["venta_total_equivalente"], Decimal("3"))

    def test_canonical_point_difference_uses_legacy_storage_sign_and_blocks_source_issue(self):
        class CanonicalBalance:
            def build(self, month, **kwargs):
                return self.balance

        canonical = CanonicalBalance()
        canonical.balance = self._canonical_balance(
            {
                self.parent.id: MonthlyPointBalanceRow(
                    receta_id=self.parent.id,
                    opening_point=Decimal("10"),
                    calculated_closing=Decimal("10"),
                    closing_point=Decimal("12"),
                    difference_point=Decimal("2"),
                    status="REVISAR_FUENTE",
                    issues=("MONTH_SOURCE_INCOMPLETE",),
                    source_counts=MappingProxyType({}),
                )
            },
            issues=("MONTH_SOURCE_INCOMPLETE",),
            selected_dates=(date(2025, 9, 29), date(2025, 9, 30)),
        )
        service = ProductMonthClosureService(balance_service=canonical)

        preview = service.preview(month="2025-09")

        row = preview["line_rows"][0]
        self.assertEqual(row["diferencia_teorico_vs_point"], Decimal("-2"))
        self.assertEqual(row["metadata"]["point_difference"], "2")
        self.assertEqual(row["estado_auditoria"], ProductoMonthClosureLine.AUDIT_STATUS_REVISAR_CATALOGO)
        self.assertFalse(preview["metadata"]["validation"]["lock_ready"])
        self.assertFalse(row["metadata"]["point_final_scopes_available"])
        self.assertEqual(
            preview["metadata"]["opening_meta"]["selected_dates"],
            ["2025-09-29", "2025-09-30"],
        )

    def test_canonical_closing_scopes_preserve_zero_and_negative_balances_and_sales_mode_alias(self):
        class CanonicalBalance:
            def __init__(self, balance):
                self.balance = balance

            def build(self, month, **kwargs):
                return self.balance

        row = MonthlyPointBalanceRow(
            receta_id=self.parent.id,
            opening_point=Decimal("0"),
            calculated_closing=Decimal("0"),
            closing_point=Decimal("0"),
            closing_point_cedis=Decimal("-2"),
            closing_point_sucursales=Decimal("2"),
            difference_point=Decimal("0"),
            status="COINCIDE",
            source_counts=MappingProxyType({"closing_snapshot_rows": 1}),
        )
        preview = ProductMonthClosureService(
            balance_service=CanonicalBalance(self._canonical_balance({self.parent.id: row}))
        ).preview(month="2025-09")

        projected = preview["line_rows"][0]
        self.assertTrue(projected["metadata"]["point_final_scopes_available"])
        self.assertEqual(projected["inventario_final_point_total"], Decimal("0"))
        self.assertEqual(projected["inventario_final_point_cedis"], Decimal("-2"))
        self.assertEqual(projected["inventario_final_point_sucursales"], Decimal("2"))
        self.assertEqual(preview["metadata"]["validation"]["closing_inventory"]["matched_recipe_count"], 1)
        self.assertEqual(preview["metadata"]["sales_meta"]["mode"], "bridge_history")

    def test_preview_counts_only_snapshot_covered_recipes_and_sums_real_scopes(self):
        class CanonicalBalance:
            def __init__(self, balance):
                self.balance = balance

            def build(self, month, **kwargs):
                return self.balance

        production_only = Receta.objects.create(
            nombre="Producto sin snapshot de cierre",
            codigo_point="SIN-SNAPSHOT",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="test-sin-snapshot",
        )
        covered = MonthlyPointBalanceRow(
            receta_id=self.parent.id,
            opening_point=Decimal("0"),
            calculated_closing=Decimal("0"),
            closing_point=Decimal("0"),
            closing_point_cedis=Decimal("0"),
            closing_point_sucursales=Decimal("0"),
            difference_point=Decimal("0"),
            status="COINCIDE",
            source_counts=MappingProxyType({"closing_snapshot_rows": 1}),
        )
        uncovered = MonthlyPointBalanceRow(
            receta_id=production_only.id,
            opening_point=Decimal("0"),
            production=Decimal("3"),
            calculated_closing=Decimal("3"),
            closing_point=None,
            difference_point=None,
            status="REVISAR_FUENTE",
            source_counts=MappingProxyType({"production_rows": 1, "closing_snapshot_rows": 0}),
        )
        preview = ProductMonthClosureService(
            balance_service=CanonicalBalance(self._canonical_balance({self.parent.id: covered, production_only.id: uncovered}))
        ).preview(month="2025-09")

        self.assertEqual(preview["metadata"]["validation"]["closing_inventory"]["matched_recipe_count"], 1)
        self.assertEqual(preview["totals"]["closing_cedis"], Decimal("0"))
        self.assertEqual(preview["totals"]["closing_sucursales"], Decimal("0"))
        self.assertEqual(preview["totals"]["closing_total"], Decimal("0"))

    def test_historical_line_without_canonical_metadata_remains_readable_without_rebuild(self):
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2025, 7, 1),
            month_end=date(2025, 7, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            metadata={},
        )
        line = ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=self.parent,
            inventario_final_teorico=Decimal("7"),
            metadata={},
        )

        loaded = ProductoMonthClosureLine.objects.get(pk=line.pk)

        self.assertEqual(loaded.inventario_final_teorico, Decimal("7"))
        self.assertEqual(loaded.metadata, {})
        self.assertEqual(ProductoMonthClosure.objects.get(pk=closure.pk).metadata, {})

    def test_build_prefers_production_facts_when_available(self):
        previous = ProductoMonthClosure.objects.create(
            month_start=date(2025, 8, 1),
            month_end=date(2025, 8, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_BOOTSTRAP_SEED,
            opening_reference_date=date(2025, 8, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=self.parent,
            inventario_final_teorico=Decimal("20"),
        )
        FactProduccionDiaria.objects.create(
            fecha=date(2025, 9, 10),
            sucursal=self.sucursal,
            receta=self.parent,
            producido=Decimal("12"),
            vendido=Decimal("8"),
            merma=Decimal("2"),
        )
        VentaHistorica.objects.create(
            receta=self.parent,
            sucursal=self.sucursal,
            fecha=date(2025, 9, 10),
            cantidad=Decimal("999"),
            fuente="POINT_BRIDGE_SALES",
        )

        closure = self.service.build(month="2025-09")

        line = closure.lines.get(receta_padre=self.parent)
        self.assertEqual(line.produccion_mes, Decimal("12"))
        self.assertEqual(line.venta_total_equivalente, Decimal("8"))
        self.assertEqual(line.merma_total_equivalente, Decimal("2"))
        self.assertEqual(line.inventario_final_teorico, Decimal("0"))
        self.assertEqual((closure.metadata or {}).get("fact_meta", {}).get("status"), "canonical")
        self.assertEqual((closure.metadata or {}).get("sales_meta", {}).get("selected_source"), "production_facts")

    def test_build_generates_production_facts_from_staging_when_missing(self):
        previous = ProductoMonthClosure.objects.create(
            month_start=date(2026, 3, 1),
            month_end=date(2026, 3, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_BOOTSTRAP_SEED,
            opening_reference_date=date(2026, 3, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=self.parent,
            inventario_final_teorico=Decimal("20"),
        )
        product = PointProduct.objects.create(
            external_id="parent-facts-1",
            sku=self.parent.codigo_point,
            name=self.parent.nombre,
            category="Pasteles",
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=product,
            receta=self.parent,
            sync_job=self.sync_job,
            sale_date=date(2026, 4, 10),
            quantity=Decimal("8"),
            tickets=0,
            gross_amount=Decimal("800"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("800"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("800"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        PointProductionLine.objects.create(
            branch=self.point_branch,
            erp_branch=self.sucursal,
            receta=self.parent,
            production_external_id="prod-facts-1",
            detail_external_id="detail-facts-1",
            source_hash="prod-facts-hash-1",
            production_date=date(2026, 4, 11),
            item_name=self.parent.nombre,
            item_code=self.parent.codigo_point,
            produced_quantity=Decimal("12"),
        )
        PointWasteLine.objects.create(
            branch=self.point_branch,
            erp_branch=self.sucursal,
            receta=self.parent,
            sync_job=self.sync_job,
            movement_external_id="waste-facts-1",
            source_hash="waste-facts-hash-1",
            movement_at=timezone.make_aware(datetime(2026, 4, 12, 12, 0)),
            item_name=self.parent.nombre,
            item_code=self.parent.codigo_point,
            quantity=Decimal("2"),
        )

        closure = self.service.build(month="2026-04")

        line = closure.lines.get(receta_padre=self.parent)
        self.assertEqual(line.produccion_mes, Decimal("12"))
        self.assertEqual(line.venta_total_equivalente, Decimal("0"))
        self.assertEqual(line.merma_total_equivalente, Decimal("2"))
        self.assertEqual(line.inventario_final_teorico, Decimal("0"))
        self.assertEqual((closure.metadata or {}).get("fact_meta", {}).get("status"), "canonical")
        self.assertFalse(FactProduccionDiaria.objects.filter(fecha=date(2026, 4, 10), receta=self.parent).exists())

    def test_build_uses_previous_closure_and_rolls_slice_sales_and_waste_to_parent(self):
        previous = ProductoMonthClosure.objects.create(
            month_start=date(2025, 8, 1),
            month_end=date(2025, 8, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
            opening_reference_date=date(2025, 7, 31),
            built_at=timezone.now(),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=self.parent,
            inventario_final_teorico=Decimal("20"),
        )

        PointProductionLine.objects.create(
            branch=self.point_branch,
            erp_branch=self.sucursal,
            receta=self.parent,
            production_external_id="prod-1",
            detail_external_id="detail-1",
            source_hash="prod-hash-1",
            production_date=date(2025, 9, 5),
            item_name=self.parent.nombre,
            item_code=self.parent.codigo_point,
            produced_quantity=Decimal("15"),
        )
        VentaHistorica.objects.create(
            receta=self.parent,
            sucursal=self.sucursal,
            fecha=date(2025, 9, 10),
            cantidad=Decimal("5"),
            fuente="POINT_BRIDGE_SALES",
        )
        VentaHistorica.objects.create(
            receta=self.derived,
            sucursal=self.sucursal,
            fecha=date(2025, 9, 11),
            cantidad=Decimal("20"),
            fuente="POINT_BRIDGE_SALES",
        )
        PointWasteLine.objects.create(
            branch=self.point_branch,
            erp_branch=self.sucursal,
            receta=self.derived,
            sync_job=self.sync_job,
            movement_external_id="waste-1",
            source_hash="waste-hash-1",
            movement_at=timezone.make_aware(datetime(2025, 9, 12, 10, 0, 0), timezone.get_current_timezone()),
            item_name=self.derived.nombre,
            quantity=Decimal("5"),
        )

        closure = self.service.build(month="2025-09")

        self.assertEqual(closure.opening_source, ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT)
        line = closure.lines.get(receta_padre=self.parent)
        self.assertEqual(line.inventario_inicial_teorico, Decimal("0"))
        self.assertEqual(line.produccion_mes, Decimal("15"))
        self.assertEqual(line.venta_directa_enteros, Decimal("5"))
        self.assertEqual(line.venta_derivada_equivalente, Decimal("2"))
        self.assertEqual(line.merma_derivada_equivalente, Decimal("0.5"))
        self.assertEqual(line.inventario_final_teorico, Decimal("0"))

    def test_build_uses_closure_equivalence_before_derived_presentations(self):
        cheesecake_parent = Receta.objects.create(
            nombre="Cheesecake Mediano",
            codigo_point="CH-M",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-cheesecake-mediano",
        )
        cheesecake_slice = Receta.objects.create(
            nombre="Cheesecake Rebanada",
            codigo_point="CH-R",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-cheesecake-rebanada",
        )
        RecetaEquivalencia.objects.create(
            receta_porcion=cheesecake_slice,
            receta_padre=cheesecake_parent,
            factor_conversion=Decimal("8"),
            fuente="test",
        )
        previous = ProductoMonthClosure.objects.create(
            month_start=date(2025, 8, 1),
            month_end=date(2025, 8, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_BOOTSTRAP_SEED,
            opening_reference_date=date(2025, 8, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=cheesecake_parent,
            inventario_final_teorico=Decimal("4"),
        )
        PointProductionLine.objects.create(
            branch=self.point_branch,
            erp_branch=self.sucursal,
            receta=cheesecake_parent,
            production_external_id="prod-cheesecake-1",
            detail_external_id="detail-cheesecake-1",
            source_hash="prod-cheesecake-hash-1",
            production_date=date(2025, 9, 5),
            item_name=cheesecake_parent.nombre,
            item_code=cheesecake_parent.codigo_point,
            produced_quantity=Decimal("6"),
        )
        VentaHistorica.objects.create(
            receta=cheesecake_slice,
            sucursal=self.sucursal,
            fecha=date(2025, 9, 11),
            cantidad=Decimal("16"),
            fuente="POINT_BRIDGE_SALES",
        )

        closure = self.service.build(month="2025-09")

        line = closure.lines.get(receta_padre=cheesecake_parent)
        self.assertEqual(line.produccion_mes, Decimal("6"))
        self.assertEqual(line.venta_derivada_equivalente, Decimal("2"))
        self.assertEqual(line.inventario_final_teorico, Decimal("0"))

    def test_build_ignores_recipes_marked_excluir_cierre(self):
        empaque = Receta.objects.create(
            nombre="Empaque Rebanada Especial",
            codigo_point="EMP-R",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            excluir_cierre=True,
            hash_contenido="hash-empaque-rebanada-especial",
        )
        previous = ProductoMonthClosure.objects.create(
            month_start=date(2025, 8, 1),
            month_end=date(2025, 8, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_BOOTSTRAP_SEED,
            opening_reference_date=date(2025, 8, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=self.parent,
            inventario_final_teorico=Decimal("5"),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=empaque,
            inventario_final_teorico=Decimal("50"),
        )
        VentaHistorica.objects.create(
            receta=empaque,
            sucursal=self.sucursal,
            fecha=date(2025, 9, 11),
            cantidad=Decimal("30"),
            fuente="POINT_BRIDGE_SALES",
        )

        with self.assertRaises(ProductMonthClosureError):
            self.service.build(month="2025-09")

    def test_build_uses_snapshot_opening_when_previous_closure_missing(self):
        point_parent = PointProduct.objects.create(external_id="point-parent", sku="SNK-M", name=self.parent.nombre)
        point_derived = PointProduct.objects.create(external_id="point-derived", sku="SNK-R", name=self.derived.nombre)

        PointInventorySnapshot.objects.create(
            branch=self.point_branch,
            product=point_parent,
            stock=Decimal("2"),
            sync_job=self.sync_job,
            captured_at=timezone.make_aware(datetime(2025, 8, 31, 23, 0, 0), timezone.get_current_timezone()),
        )
        PointInventorySnapshot.objects.create(
            branch=self.point_branch,
            product=point_derived,
            stock=Decimal("5"),
            sync_job=self.sync_job,
            captured_at=timezone.make_aware(datetime(2025, 8, 31, 23, 0, 0), timezone.get_current_timezone()),
        )

        closure = self.service.build(month="2025-09")

        self.assertEqual(closure.opening_source, ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT)
        line = closure.lines.get(receta_padre=self.parent)
        self.assertEqual(line.inventario_inicial_teorico, Decimal("2.5"))
        self.assertEqual(line.source_snapshot_count, 2)
        self.assertEqual(line.inventario_final_teorico, Decimal("0"))
        self.assertIn("CALCULATED_CLOSING_MISSING", line.metadata["issues"])

    def test_build_rejects_locked_closure_rebuild(self):
        ProductoMonthClosure.objects.create(
            month_start=date(2025, 9, 1),
            month_end=date(2025, 9, 30),
            status=ProductoMonthClosure.STATUS_LOCKED,
            is_locked=True,
        )
        with self.assertRaisesMessage(ProductMonthClosureError, "bloqueado"):
            self.service.build(month="2025-09", rebuild=True)

    def test_build_rejects_existing_closure_without_explicit_rebuild_and_preserves_it(self):
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2025, 9, 1),
            month_end=date(2025, 9, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            metadata={"sentinel": "preserve"},
        )
        line = ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=self.parent,
            inventario_final_teorico=Decimal("17"),
        )

        with patch.object(self.service, "preview") as preview:
            with self.assertRaisesMessage(ProductMonthClosureError, "se requiere rebuild"):
                self.service.build(month="2025-09")

        preview.assert_not_called()
        closure.refresh_from_db()
        line.refresh_from_db()
        self.assertEqual(closure.metadata, {"sentinel": "preserve"})
        self.assertEqual(line.inventario_final_teorico, Decimal("17"))

    def test_bootstrap_rejects_existing_closure_without_explicit_rebuild_and_preserves_it(self):
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2026, 6, 1),
            month_end=date(2026, 6, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            metadata={"sentinel": "preserve-bootstrap"},
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=self.parent,
            inventario_final_teorico=Decimal("9"),
        )

        with patch.object(self.service, "_build_bootstrap_seed_rows") as projector:
            with self.assertRaisesMessage(ProductMonthClosureError, "se requiere rebuild"):
                self.service.build_bootstrap_seed(
                    month="2026-06",
                    seed_rows=[],
                    source_label="histórico",
                )

        projector.assert_not_called()
        closure.refresh_from_db()
        self.assertEqual(closure.metadata, {"sentinel": "preserve-bootstrap"})
        self.assertEqual(closure.lines.get().inventario_final_teorico, Decimal("9"))

    def test_build_revalidates_locked_closure_after_preview_and_never_unlocks_it(self):
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2025, 9, 1),
            month_end=date(2025, 9, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            is_locked=False,
            metadata={"sentinel": "before-lock"},
        )
        original_updated_at = closure.updated_at

        def preview_with_concurrent_lock(**kwargs):
            ProductoMonthClosure.objects.filter(pk=closure.pk).update(
                status=ProductoMonthClosure.STATUS_LOCKED,
                is_locked=True,
                metadata={"sentinel": "concurrent-lock"},
            )
            return {
                "month_end": date(2025, 9, 30),
                "notes": "plan obsoleto",
                "opening_source": ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
                "opening_reference_date": date(2025, 8, 31),
                "metadata": {},
                "line_rows": [],
            }

        with patch.object(self.service, "preview", side_effect=preview_with_concurrent_lock):
            with self.assertRaisesMessage(ProductMonthClosureError, "bloqueado"):
                self.service.build(month="2025-09", rebuild=True)

        closure.refresh_from_db()
        # El callback corre dentro de la misma transacción serializada; al
        # detectar el cambio, todo el intento (incluido ese update simulado)
        # se revierte y el cierre previo queda intacto.
        self.assertFalse(closure.is_locked)
        self.assertEqual(closure.status, ProductoMonthClosure.STATUS_BUILT)
        self.assertEqual(closure.metadata, {"sentinel": "before-lock"})
        self.assertEqual(closure.updated_at, original_updated_at)

    def test_lock_reloads_and_rejects_a_stale_instance_locked_by_another_request(self):
        stale = ProductoMonthClosure.objects.create(
            month_start=date(2026, 5, 1),
            month_end=date(2026, 5, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            is_locked=False,
            metadata={"sentinel": "stale"},
        )
        ProductoMonthClosureLine.objects.create(
            closure=stale,
            receta_padre=self.parent,
        )
        ProductoMonthClosure.objects.filter(pk=stale.pk).update(
            status=ProductoMonthClosure.STATUS_LOCKED,
            is_locked=True,
            metadata={"sentinel": "concurrent-lock"},
        )

        with self.assertRaisesMessage(ProductMonthClosureError, "ya esta bloqueado"):
            self.service.lock(closure=stale, reason="stale-request")

        stale.refresh_from_db()
        self.assertTrue(stale.is_locked)
        self.assertEqual(stale.status, ProductoMonthClosure.STATUS_LOCKED)
        self.assertEqual(stale.metadata, {"sentinel": "concurrent-lock"})

    def test_bootstrap_seed_revalidates_concurrent_lock_before_rebuild(self):
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2026, 6, 1),
            month_end=date(2026, 6, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            is_locked=False,
            metadata={"sentinel": "before-lock"},
        )
        original_updated_at = closure.updated_at

        def seed_rows_with_concurrent_lock(**kwargs):
            ProductoMonthClosure.objects.filter(pk=closure.pk).update(
                status=ProductoMonthClosure.STATUS_LOCKED,
                is_locked=True,
                metadata={"sentinel": "concurrent-lock"},
            )
            return [], {}, {}

        with patch.object(
            self.service,
            "_build_bootstrap_seed_rows",
            side_effect=seed_rows_with_concurrent_lock,
        ):
            with self.assertRaisesMessage(ProductMonthClosureError, "bloqueado"):
                self.service.build_bootstrap_seed(
                    month="2026-06",
                    seed_rows=[],
                    source_label="cierre histórico",
                    rebuild=True,
                )

        closure.refresh_from_db()
        self.assertFalse(closure.is_locked)
        self.assertEqual(closure.status, ProductoMonthClosure.STATUS_BUILT)
        self.assertEqual(closure.metadata, {"sentinel": "before-lock"})
        self.assertEqual(closure.updated_at, original_updated_at)

    def test_historical_excel_import_revalidates_concurrent_lock_before_rebuild(self):
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2026, 7, 1),
            month_end=date(2026, 7, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            is_locked=False,
            metadata={"sentinel": "before-lock"},
        )
        original_updated_at = closure.updated_at
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "JULIO 26"
        worksheet["B1"] = self.parent.nombre
        worksheet["D1"] = 5
        worksheet["I1"] = 1
        worksheet["J1"] = 2
        worksheet["K1"] = 3

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            workbook.save(tmp.name)
            concurrently_locked = ProductoMonthClosure.objects.get(pk=closure.pk)
            concurrently_locked.status = ProductoMonthClosure.STATUS_LOCKED
            concurrently_locked.is_locked = True
            concurrently_locked.metadata = {"sentinel": "concurrent-lock"}

            with (
                patch(
                    "pos_bridge.management.commands.import_historical_product_closure_excel._resolve_receta",
                    return_value=self.parent,
                ),
                patch(
                    "pos_bridge.management.commands.import_historical_product_closure_excel._sales_map",
                    return_value=({}, "test"),
                ),
                patch(
                    "pos_bridge.management.commands.import_historical_product_closure_excel._production_map",
                    return_value=({}, "test"),
                ),
                patch(
                    "pos_bridge.management.commands.import_historical_product_closure_excel._merma_maps",
                    return_value=({}, {}, "test"),
                ),
                patch.object(
                    ProductoMonthClosure.objects,
                    "select_for_update",
                ) as select_for_update,
            ):
                select_for_update.return_value.get.return_value = concurrently_locked
                with self.assertRaisesMessage(CommandError, "bloqueado"):
                    call_command(
                        "import_historical_product_closure_excel",
                        tmp.name,
                        sheet="JULIO 26",
                        month="2026-07",
                        rebuild=True,
                    )
                select_for_update.assert_called_once_with()

        closure.refresh_from_db()
        self.assertFalse(closure.is_locked)
        self.assertEqual(closure.status, ProductoMonthClosure.STATUS_BUILT)
        self.assertEqual(closure.metadata, {"sentinel": "before-lock"})
        self.assertEqual(closure.updated_at, original_updated_at)

    def test_historical_excel_import_metadata_names_physical_count_without_claiming_point(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "JUNIO 25"
        worksheet["B1"] = self.parent.nombre
        worksheet["D1"] = 5
        worksheet["I1"] = 1
        worksheet["J1"] = 2
        worksheet["K1"] = 3

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            workbook.save(tmp.name)
            with (
                patch(
                    "pos_bridge.management.commands.import_historical_product_closure_excel._resolve_receta",
                    return_value=self.parent,
                ),
                patch(
                    "pos_bridge.management.commands.import_historical_product_closure_excel._sales_map",
                    return_value=({}, "test"),
                ),
                patch(
                    "pos_bridge.management.commands.import_historical_product_closure_excel._production_map",
                    return_value=({}, "test"),
                ),
                patch(
                    "pos_bridge.management.commands.import_historical_product_closure_excel._merma_maps",
                    return_value=({}, {}, "test"),
                ),
            ):
                call_command(
                    "import_historical_product_closure_excel",
                    tmp.name,
                    sheet="JUNIO 25",
                    month="2025-06",
                    stdout=StringIO(),
                )

        closure = ProductoMonthClosure.objects.get(month_start=date(2025, 6, 1))
        imported_columns = closure.metadata["historical_excel_import"]["imported_columns"]
        self.assertEqual(
            imported_columns,
            [
                "inventario_inicial_historico",
                "conteo_historico_cedis",
                "conteo_historico_sucursales",
                "inventario_historico_fisico_total",
            ],
        )
        historical = closure.lines.get().metadata["historical_excel"]
        self.assertEqual(historical["conteo_historico_cedis"], "1")
        self.assertEqual(historical["conteo_historico_sucursales"], "2")
        self.assertEqual(historical["inventario_historico_fisico_total"], "3")
        self.assertFalse(any("point" in key.lower() for key in historical))
        self.assertEqual(
            historical["movement_authority"],
            {
                "sales": {"source": "test", "source_present": False, "authoritative": False},
                "production": {"source": "test", "source_present": False, "authoritative": False},
                "waste": {"source": "test", "source_present": False, "authoritative": False},
                "conversions": {"source": "sin_datos", "source_present": False, "authoritative": False},
            },
        )
        from pos_bridge.services.product_closure_projection import project_product_closure_line

        projected = project_product_closure_line(closure.lines.get(), historical_excel_import=True)
        for field in ("production", "sales_total", "waste_total", "point_conversion_in", "calculated_closing"):
            self.assertIsNone(projected[field], field)

    def test_historical_excel_preserves_blank_scopes_and_explicit_zero_total(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "MAYO 25"
        worksheet["B1"] = self.parent.nombre
        worksheet["K1"] = 0

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            workbook.save(tmp.name)
            with (
                patch(
                    "pos_bridge.management.commands.import_historical_product_closure_excel._resolve_receta",
                    return_value=self.parent,
                ),
                patch(
                    "pos_bridge.management.commands.import_historical_product_closure_excel._sales_map",
                    return_value=({}, "test"),
                ),
                patch(
                    "pos_bridge.management.commands.import_historical_product_closure_excel._production_map",
                    return_value=({}, "test"),
                ),
                patch(
                    "pos_bridge.management.commands.import_historical_product_closure_excel._merma_maps",
                    return_value=({}, {}, "test"),
                ),
            ):
                call_command(
                    "import_historical_product_closure_excel",
                    tmp.name,
                    sheet="MAYO 25",
                    month="2025-05",
                    stdout=StringIO(),
                )

        line = ProductoMonthClosure.objects.get(month_start=date(2025, 5, 1)).lines.get()
        historical = line.metadata["historical_excel"]
        self.assertEqual(
            historical["inventory_presence"],
            {"opening": False, "cedis": False, "sucursales": False, "total": True},
        )
        from pos_bridge.services.product_closure_projection import project_product_closure_line

        projected = project_product_closure_line(line, historical_excel_import=True)
        self.assertIsNone(projected["historical_opening"])
        self.assertIsNone(projected["opening_balance"])
        self.assertIsNone(projected["historical_count_cedis"])
        self.assertIsNone(projected["historical_count_sucursales"])
        self.assertEqual(projected["historical_count"], Decimal("0"))

    def test_legacy_historical_presence_is_inferred_only_from_metadata_keys(self):
        line = ProductoMonthClosureLine.objects.create(
            closure=ProductoMonthClosure.objects.create(
                month_start=date(2025, 4, 1),
                month_end=date(2025, 4, 30),
                status=ProductoMonthClosure.STATUS_BUILT,
            ),
            receta_padre=self.parent,
            inventario_inicial_teorico=Decimal("0"),
            inventario_final_point_cedis=Decimal("0"),
            inventario_final_point_sucursales=Decimal("7"),
            inventario_final_point_total=Decimal("0"),
            metadata={"historical_excel": {
                "inventario_inicial_historico": "0",
                "conteo_historico_cedis": "0",
                "inventario_historico_fisico_total": "0",
            }},
        )
        from pos_bridge.services.product_closure_projection import project_product_closure_line

        projected = project_product_closure_line(line, historical_excel_import=True)
        self.assertIsNone(projected["opening_point"])
        self.assertEqual(projected["historical_opening"], Decimal("0"))
        self.assertEqual(projected["historical_count_cedis"], Decimal("0"))
        self.assertIsNone(projected["historical_count_sucursales"])
        self.assertEqual(projected["historical_count"], Decimal("0"))

    def test_lock_marks_built_closure_as_locked_with_audit_metadata(self):
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2025, 9, 1),
            month_end=date(2025, 9, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
            opening_reference_date=date(2025, 8, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=self.parent,
            inventario_inicial_teorico=Decimal("5"),
            inventario_final_teorico=Decimal("5"),
        )

        locked = self.service.lock(closure=closure)

        self.assertTrue(locked.is_locked)
        self.assertEqual(locked.status, ProductoMonthClosure.STATUS_LOCKED)
        self.assertIn("lock_event", locked.metadata)
        self.assertEqual(locked.metadata["lock_event"]["line_count"], 1)

    def test_canonical_build_persists_source_fingerprint_and_unchanged_lock_succeeds(self):
        class MutableBalanceService:
            def __init__(self, balance):
                self.balance = balance
                self.refresh_values = []

            def build(self, month, **kwargs):
                self.refresh_values.append(kwargs.get("refresh_official_sales"))
                return self.balance

        balance_service = MutableBalanceService(self._lockable_fingerprint_balance())
        service = ProductMonthClosureService(balance_service=balance_service)

        closure = service.build(month="2025-09")

        fingerprint = closure.metadata.get("source_fingerprint")
        self.assertEqual(fingerprint["algorithm"], "sha256")
        self.assertEqual(len(fingerprint["digest"]), 64)
        self.assertEqual(len(fingerprint["metadata_digest"]), 64)
        self.assertEqual(len(fingerprint["projected_lines_digest"]), 64)
        self.assertEqual(len(fingerprint["raw_sources_digest"]), 64)
        with patch.object(
            service,
            "_fresh_canonical_preview",
            side_effect=lambda *, month: service.preview(month=month, refresh_official_sales=False),
        ):
            locked = service.lock(closure=closure, reason="fuentes sin cambios")
        self.assertTrue(locked.is_locked)
        self.assertIs(balance_service.refresh_values[-1], False)

    def test_canonical_lock_rejects_tampered_persisted_line(self):
        class BalanceService:
            def __init__(self, balance):
                self.balance = balance

            def build(self, month, **kwargs):
                return self.balance

        service = ProductMonthClosureService(
            balance_service=BalanceService(self._lockable_fingerprint_balance())
        )
        closure = service.build(month="2025-09")
        closure.lines.update(venta_total_equivalente=Decimal("999"))

        with self.assertRaisesMessage(ProductMonthClosureError, "líneas persistidas cambiaron"):
            service.lock(closure=closure)

        closure.refresh_from_db()
        self.assertFalse(closure.is_locked)

    def test_line_digest_covers_metadata_identity_and_distinguishes_none_from_zero(self):
        class BalanceService:
            def build(_self, month, **kwargs):
                return self._lockable_fingerprint_balance()

        service = ProductMonthClosureService(balance_service=BalanceService())
        plan = service.preview(month="2025-09")
        baseline = plan["metadata"]["source_fingerprint"]["projected_lines_digest"]
        row = {**plan["line_rows"][0], "metadata": dict(plan["line_rows"][0]["metadata"])}

        row["metadata"]["issues"] = ["TAMPERED"]
        metadata_digest = service._projected_lines_digest(
            [row],
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
        )
        self.assertNotEqual(metadata_digest, baseline)

        row = {**plan["line_rows"][0], "receta": self.derived}
        identity_digest = service._projected_lines_digest(
            [row],
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
        )
        self.assertNotEqual(identity_digest, baseline)

        row = {**plan["line_rows"][0], "inventario_final_point_total": None}
        none_digest = service._projected_lines_digest(
            [row],
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
        )
        row["inventario_final_point_total"] = Decimal("0")
        zero_digest = service._projected_lines_digest(
            [row],
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
        )
        self.assertNotEqual(none_digest, zero_digest)

    def test_compensating_raw_sales_edits_change_fingerprint_even_when_total_is_equal(self):
        product = PointProduct.objects.create(external_id="raw-sales", sku="RAW", name="Raw")
        first = PointDailySale.objects.create(
            branch=self.point_branch,
            product=product,
            receta=self.parent,
            sale_date=date(2025, 9, 10),
            quantity=Decimal("4"),
        )
        second = PointDailySale.objects.create(
            branch=self.point_branch,
            product=product,
            receta=self.parent,
            sale_date=date(2025, 9, 11),
            quantity=Decimal("6"),
        )

        class BalanceService:
            def build(_self, month, **kwargs):
                return self._lockable_fingerprint_balance()

        service = ProductMonthClosureService(balance_service=BalanceService())
        closure = service.build(month="2025-09")
        fingerprint = closure.metadata["source_fingerprint"]
        original_digest = fingerprint["raw_sources_digest"]
        raw_sales = ProductMonthClosureService._raw_source_evidence(
            month_start=date(2025, 9, 1)
        )["sales_point_daily"]
        self.assertEqual(raw_sales["row_count"], 2)
        self.assertTrue(
            {"id", "branch_id", "product_id", "sync_job_id", "sale_date", "quantity", "updated_at"}
            <= set(raw_sales["fields"])
        )

        PointDailySale.objects.filter(pk=first.pk).update(quantity=Decimal("5"))
        PointDailySale.objects.filter(pk=second.pk).update(quantity=Decimal("5"))

        current = service.preview(month="2025-09", refresh_official_sales=False)
        self.assertNotEqual(
            current["metadata"]["source_fingerprint"]["raw_sources_digest"],
            original_digest,
        )

    def test_fresh_canonical_preview_does_not_reuse_cached_matcher_or_balance_service(self):
        class StaleBalanceService:
            def build(self, month, **kwargs):
                raise AssertionError("must not reuse the service cached at build time")

        service = ProductMonthClosureService(balance_service=StaleBalanceService())
        service.matcher._point_code_index_built = True
        service.matcher._point_code_to_receta = {"stale": self.parent}
        balance = self._lockable_fingerprint_balance()
        fresh_matcher = type(service.matcher)()

        with (
            patch(
                "pos_bridge.services.product_month_closure_service.PointSalesMatchingService",
                return_value=fresh_matcher,
            ) as matcher_factory,
            patch.object(MonthlyPointProductBalanceService, "build", return_value=balance),
        ):
            plan = service._fresh_canonical_preview(month=date(2025, 9, 1))

        self.assertEqual(plan["month_start"], date(2025, 9, 1))
        matcher_factory.assert_called_once_with()
        self.assertIsNot(service.matcher, fresh_matcher)
        self.assertFalse(fresh_matcher._point_code_index_built)

    def test_canonical_lock_rejects_when_current_source_fingerprint_changed(self):
        class MutableBalanceService:
            def __init__(self, balance):
                self.balance = balance

            def build(self, month, **kwargs):
                return self.balance

        balance_service = MutableBalanceService(self._lockable_fingerprint_balance(sales_job_id=101))
        service = ProductMonthClosureService(balance_service=balance_service)
        closure = service.build(month="2025-09")
        balance_service.balance = self._lockable_fingerprint_balance(sales_job_id=202)

        with patch.object(
            service,
            "_fresh_canonical_preview",
            side_effect=lambda *, month: service.preview(month=month, refresh_official_sales=False),
        ):
            with self.assertRaisesMessage(
                ProductMonthClosureError,
                "Las fuentes cambiaron; reconstruye el cierre antes de bloquearlo.",
            ):
                service.lock(closure=closure, reason="fuente nueva")

        closure.refresh_from_db()
        self.assertFalse(closure.is_locked)
        self.assertEqual(closure.status, ProductoMonthClosure.STATUS_BUILT)

    def test_canonical_lock_rejects_missing_or_tampered_source_fingerprint(self):
        class BalanceService:
            def __init__(self, balance):
                self.balance = balance

            def build(self, month, **kwargs):
                return self.balance

        service = ProductMonthClosureService(
            balance_service=BalanceService(self._lockable_fingerprint_balance())
        )
        closure = service.build(month="2025-09")
        metadata = dict(closure.metadata)
        metadata.pop("source_fingerprint")
        closure.metadata = metadata
        closure.save(update_fields=["metadata", "updated_at"])

        with self.assertRaisesMessage(ProductMonthClosureError, "no tiene huella de fuentes"):
            service.lock(closure=closure)

        closure = service.build(month="2025-09", rebuild=True)
        metadata = dict(closure.metadata)
        metadata["sales_meta"] = {**metadata["sales_meta"], "job_id": 999}
        closure.metadata = metadata
        closure.save(update_fields=["metadata", "updated_at"])
        with self.assertRaisesMessage(
            ProductMonthClosureError,
            "La metadata de fuentes del cierre cambió; reconstruye el cierre antes de bloquearlo.",
        ):
            service.lock(closure=closure)

    def test_historical_excel_closure_keeps_explicit_lock_policy_without_point_fingerprint(self):
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2025, 7, 1),
            month_end=date(2025, 7, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            metadata={"historical_excel_import": {"scope": "inventory_only"}},
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=self.parent,
            metadata={"historical_excel": {"scope": "inventory_only"}},
        )

        locked = self.service.lock(closure=closure, reason="cierre histórico documentado")

        self.assertTrue(locked.is_locked)
        self.assertNotIn("source_fingerprint", locked.metadata)

    def test_lock_rejects_validation_not_ready_without_blocking_issues(self):
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2026, 4, 1),
            month_end=date(2026, 4, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
            metadata={
                "validation": {"lock_ready": False, "blocking_issues": []},
                "balance": {"contract": "POINT_PRODUCT_BALANCE_V1"},
                "opening_meta": {"authoritative": True, "source_present": True},
                "sales_meta": {"authoritative": True, "source_present": True},
                "production_meta": {"authoritative": True, "source_present": True},
                "waste_meta": {"authoritative": True, "source_present": True},
                "conversion_meta": {"authoritative": True, "source_present": True},
                "closing_inventory_meta": {"authoritative": True, "source_present": True},
            },
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=self.parent,
            metadata={"balance_contract": "POINT_PRODUCT_BALANCE_V1"},
        )

        with self.assertRaisesMessage(ProductMonthClosureError, "no esta listo para bloquearse"):
            self.service.lock(closure=closure)

    def test_lock_rejects_each_required_canonical_source_without_authority(self):
        required_sources = (
            "opening_meta",
            "sales_meta",
            "production_meta",
            "waste_meta",
            "conversion_meta",
            "closing_inventory_meta",
        )
        case_index = 0
        for source_key in required_sources:
            for source_state in ("absent", "non_authoritative"):
                case_index += 1
                with self.subTest(source_key=source_key, source_state=source_state):
                    month = ((case_index - 1) % 12) + 1
                    year = 2027 + ((case_index - 1) // 12)
                    sources = {
                        key: {"authoritative": True, "source_present": True}
                        for key in required_sources
                    }
                    if source_state == "absent":
                        sources.pop(source_key)
                    else:
                        sources[source_key] = {"authoritative": False, "source_present": False}
                    closure = ProductoMonthClosure.objects.create(
                        month_start=date(year, month, 1),
                        month_end=date(year, month, 28),
                        status=ProductoMonthClosure.STATUS_BUILT,
                        opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
                        metadata={
                            "validation": {"lock_ready": True, "blocking_issues": []},
                            "balance": {"contract": "POINT_PRODUCT_BALANCE_V1"},
                            **sources,
                        },
                    )
                    ProductoMonthClosureLine.objects.create(
                        closure=closure,
                        receta_padre=self.parent,
                        metadata={"balance_contract": "POINT_PRODUCT_BALANCE_V1"},
                    )

                    with self.assertRaisesMessage(ProductMonthClosureError, "fuente requerida"):
                        self.service.lock(closure=closure)

    def test_lock_rejects_closure_with_catalog_issues(self):
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2025, 9, 1),
            month_end=date(2025, 9, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
            opening_reference_date=date(2025, 8, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=self.parent,
            inventario_inicial_teorico=Decimal("5"),
            inventario_final_teorico=Decimal("5"),
            has_catalog_issue=True,
            catalog_issue_note="Relacion derivada faltante",
        )

        with self.assertRaisesMessage(ProductMonthClosureError, "incidencias de catalogo"):
            self.service.lock(closure=closure)

    @override_settings(PRODUCT_MONTH_CLOSURE_SNAPSHOT_TOLERANCE_DAYS=3)
    def test_preview_uses_snapshot_fallback_within_tolerance_and_marks_warning(self):
        point_parent = PointProduct.objects.create(external_id="point-parent-fallback", sku="SNK-M", name=self.parent.nombre)
        PointInventorySnapshot.objects.create(
            branch=self.point_branch,
            product=point_parent,
            stock=Decimal("4"),
            sync_job=self.sync_job,
            captured_at=timezone.make_aware(datetime(2025, 8, 29, 12, 0, 0), timezone.get_current_timezone()),
        )

        preview = self.service.preview(month="2025-09")

        self.assertEqual(preview["opening_reference_date"], date(2025, 8, 29))
        self.assertIn("2025-08-29", preview["metadata"]["opening_meta"]["selected_dates"])
        self.assertFalse(preview["metadata"]["validation"]["lock_ready"])

    def test_lock_rejects_closure_with_unmatched_opening_products(self):
        closure = ProductoMonthClosure.objects.create(
            month_start=date(2025, 9, 1),
            month_end=date(2025, 9, 30),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT,
            opening_reference_date=date(2025, 8, 31),
            metadata={
                "opening_meta": {"unmatched_products": ["Producto Sin Match"]},
                "validation": {"blocking_issues": ["Productos sin homologacion"]},
            },
        )
        ProductoMonthClosureLine.objects.create(
            closure=closure,
            receta_padre=self.parent,
            inventario_inicial_teorico=Decimal("5"),
            inventario_final_teorico=Decimal("5"),
        )

        with self.assertRaisesMessage(ProductMonthClosureError, "opening sin homologacion"):
            self.service.lock(closure=closure)

    def test_build_chain_uses_previous_month_closure_as_next_opening(self):
        point_parent = PointProduct.objects.create(external_id="point-parent-chain", sku="SNK-M", name=self.parent.nombre)
        PointInventorySnapshot.objects.create(
            branch=self.point_branch,
            product=point_parent,
            stock=Decimal("10"),
            sync_job=self.sync_job,
            captured_at=timezone.make_aware(datetime(2025, 8, 31, 23, 0, 0), timezone.get_current_timezone()),
        )
        september = self.service.build(month="2025-09")
        september_line = september.lines.get(receta_padre=self.parent)
        self.assertEqual(september_line.inventario_final_teorico, Decimal("0"))
        self.assertIn("CALCULATED_CLOSING_MISSING", september_line.metadata["issues"])

        PointProductionLine.objects.create(
            branch=self.point_branch,
            erp_branch=self.sucursal,
            receta=self.parent,
            production_external_id="prod-chain-1",
            detail_external_id="detail-chain-1",
            source_hash="prod-chain-hash-1",
            production_date=date(2025, 10, 2),
            item_name=self.parent.nombre,
            item_code=self.parent.codigo_point,
            produced_quantity=Decimal("3"),
        )

        october = self.service.build(month="2025-10")
        october_line = october.lines.get(receta_padre=self.parent)
        self.assertEqual(october.opening_source, ProductoMonthClosure.OPENING_SOURCE_POINT_SNAPSHOT)
        self.assertEqual(october_line.inventario_inicial_teorico, Decimal("0"))
        self.assertEqual(october_line.inventario_final_teorico, Decimal("0"))

    def test_backfill_command_dry_run_reports_month_summary(self):
        point_parent = PointProduct.objects.create(external_id="point-parent-backfill", sku="SNK-M", name=self.parent.nombre)
        PointInventorySnapshot.objects.create(
            branch=self.point_branch,
            product=point_parent,
            stock=Decimal("7"),
            sync_job=self.sync_job,
            captured_at=timezone.make_aware(datetime(2025, 8, 31, 23, 0, 0), timezone.get_current_timezone()),
        )

        out = StringIO()
        call_command(
            "backfill_product_month_closure",
            from_month="2025-09",
            to_month="2025-09",
            dry_run=True,
            stdout=out,
        )

        payload = out.getvalue()
        self.assertIn('"dry_run": true', payload)
        self.assertIn('"month": "2025-09"', payload)
        self.assertIn('"status": "warning"', payload)
        self.assertIn('"lock_ready": false', payload)

    def test_cargar_equivalencias_porciones_command_is_idempotent(self):
        cheesecake_parent = Receta.objects.create(
            nombre="Cheesecake Mediano Command",
            codigo_point="CH-M-C",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-cheesecake-mediano-command",
        )
        cheesecake_slice = Receta.objects.create(
            nombre="Cheesecake Rebanada Command",
            codigo_point="CH-R-C",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-cheesecake-rebanada-command",
        )
        empaque = Receta.objects.create(
            nombre="Empaque Command",
            codigo_point="EMP-C",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-empaque-command",
        )
        csv_payload = (
            "receta_porcion,receta_padre_confirmada,factor_conversion,confirmado\n"
            f"{cheesecake_slice.nombre},{cheesecake_parent.nombre},8,SI\n"
            f"{empaque.nombre},{empaque.nombre},1,EXCLUIR\n"
        )

        with NamedTemporaryFile(mode="w+", suffix=".csv", encoding="utf-8") as tmp:
            tmp.write(csv_payload)
            tmp.flush()
            call_command("cargar_equivalencias_porciones", archivo=tmp.name, ejecutar=True, stdout=StringIO())
            call_command("cargar_equivalencias_porciones", archivo=tmp.name, ejecutar=True, stdout=StringIO())

        equivalence = RecetaEquivalencia.objects.get(receta_porcion=cheesecake_slice)
        empaque.refresh_from_db()
        self.assertEqual(equivalence.receta_padre, cheesecake_parent)
        self.assertEqual(equivalence.factor_conversion, Decimal("8.000000"))
        self.assertTrue(empaque.excluir_cierre)
        self.assertEqual(RecetaEquivalencia.objects.filter(receta_porcion=cheesecake_slice).count(), 1)

    def test_build_carries_forward_previous_opening_issues(self):
        previous = ProductoMonthClosure.objects.create(
            month_start=date(2025, 8, 1),
            month_end=date(2025, 8, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_BOOTSTRAP_SEED,
            opening_reference_date=date(2025, 8, 31),
            metadata={
                "opening_meta": {"bootstrap_seeded": True, "unmatched_products": ["Producto Sin Match"]},
                "validation": {"blocking_issues": ["Productos sin homologacion"]},
                "bootstrap_seed": {"is_seed": True, "source_label": "bootstrap.xlsx::SEPT 25::D"},
            },
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=self.parent,
            inventario_inicial_teorico=Decimal("8"),
            inventario_final_teorico=Decimal("8"),
        )

        with self.assertRaises(ProductMonthClosureError):
            self.service.build(month="2025-09")

    def test_bootstrap_command_builds_seed_closure_from_excel(self):
        out = StringIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "SEPT 25"
        ws["A1"] = "PRODUCTO"
        ws["D1"] = "INVENTARIO INICIAL"
        ws["A2"] = self.parent.nombre
        ws["D2"] = 2
        ws["A3"] = self.derived.nombre
        ws["D3"] = 5
        ws["A4"] = "PRODUCTO / REBANADAS"
        ws["D4"] = ""

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            call_command(
                "bootstrap_product_month_closure",
                tmp.name,
                sheet="SEPT 25",
                seed_month="2025-08",
                name_column="A",
                stdout=out,
            )

        closure = ProductoMonthClosure.objects.get(month_start=date(2025, 8, 1))
        self.assertEqual(closure.opening_source, ProductoMonthClosure.OPENING_SOURCE_BOOTSTRAP_SEED)
        self.assertEqual(closure.status, ProductoMonthClosure.STATUS_BUILT)
        line = closure.lines.get(receta_padre=self.parent)
        self.assertEqual(line.inventario_inicial_teorico, Decimal("2"))
        self.assertEqual(line.inventario_final_teorico, Decimal("2"))
        validation = (closure.metadata or {}).get("validation", {})
        self.assertTrue(validation["bootstrap_seeded"])
        self.assertEqual(validation["unmatched_opening_products_count"], 0)
        opening_meta = (closure.metadata or {}).get("opening_meta", {})
        self.assertEqual(opening_meta.get("derived_rows_ignored"), 1)
        self.assertIn('"opening_source": "BOOTSTRAP_SEED"', out.getvalue())

    def test_build_excludes_preparations_vasos_and_accessory_like_products(self):
        vaso = Receta.objects.create(
            nombre="Vaso Fresas con Crema Mini",
            codigo_point="VASO-FCM",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Vaso Preparado Mini",
            familia="Vasos Preparados",
            hash_contenido="hash-vaso-fcm",
        )
        letrero = Receta.objects.create(
            nombre="Letrero Chispas Felicidades",
            codigo_point="LETRERO-1",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-letrero-1",
        )
        preparacion = Receta.objects.create(
            nombre="Crema pastelera",
            codigo_point="CREMA-1",
            tipo=Receta.TIPO_PREPARACION,
            categoria="Betún, Cremas, Rellenos (INSUMO PRODUCIDO)",
            familia="Betún, Cremas, Rellenos (INSUMO PRODUCIDO)",
            hash_contenido="hash-crema-1",
        )

        previous = ProductoMonthClosure.objects.create(
            month_start=date(2025, 8, 1),
            month_end=date(2025, 8, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_BOOTSTRAP_SEED,
            opening_reference_date=date(2025, 8, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=self.parent,
            inventario_inicial_teorico=Decimal("10"),
            inventario_final_teorico=Decimal("10"),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=vaso,
            inventario_inicial_teorico=Decimal("3"),
            inventario_final_teorico=Decimal("3"),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=letrero,
            inventario_inicial_teorico=Decimal("2"),
            inventario_final_teorico=Decimal("2"),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=preparacion,
            inventario_inicial_teorico=Decimal("4"),
            inventario_final_teorico=Decimal("4"),
        )

        with self.assertRaises(ProductMonthClosureError):
            self.service.build(month="2025-09")

    def test_build_excludes_kg_and_sabor_modifier_products(self):
        kg_recipe = Receta.objects.create(
            nombre="Bolitas de Nuez KG",
            codigo_point="05021",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Galletas",
            familia="Galletas",
            hash_contenido="hash-bolitas-kg",
        )
        sabor_recipe = Receta.objects.create(
            nombre="Sabor Fresa Grande Pay",
            codigo_point="SFRESAG",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Pay Grande",
            familia="Pay",
            hash_contenido="hash-sabor-fresa-pay",
        )

        previous = ProductoMonthClosure.objects.create(
            month_start=date(2025, 8, 1),
            month_end=date(2025, 8, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_BOOTSTRAP_SEED,
            opening_reference_date=date(2025, 8, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=self.parent,
            inventario_inicial_teorico=Decimal("10"),
            inventario_final_teorico=Decimal("10"),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=kg_recipe,
            inventario_inicial_teorico=Decimal("3"),
            inventario_final_teorico=Decimal("3"),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=sabor_recipe,
            inventario_inicial_teorico=Decimal("5"),
            inventario_final_teorico=Decimal("5"),
        )

        with self.assertRaises(ProductMonthClosureError):
            self.service.build(month="2025-09")

    def test_build_excludes_topping_and_sin_preparar_products(self):
        topping_recipe = Receta.objects.create(
            nombre="TOPPING FRESA M",
            codigo_point="SFRESAPM",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            categoria="Pastel Mediano",
            hash_contenido="hash-topping-fresa-m",
        )
        sin_preparar_recipe = Receta.objects.create(
            nombre="Pan de Muerto Sin Preparar",
            codigo_point="0124",
            tipo=Receta.TIPO_PRODUCTO_FINAL,
            hash_contenido="hash-pan-muerto-sin-preparar",
        )

        previous = ProductoMonthClosure.objects.create(
            month_start=date(2025, 8, 1),
            month_end=date(2025, 8, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_BOOTSTRAP_SEED,
            opening_reference_date=date(2025, 8, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=self.parent,
            inventario_inicial_teorico=Decimal("10"),
            inventario_final_teorico=Decimal("10"),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=topping_recipe,
            inventario_inicial_teorico=Decimal("8"),
            inventario_final_teorico=Decimal("8"),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=sin_preparar_recipe,
            inventario_inicial_teorico=Decimal("4"),
            inventario_final_teorico=Decimal("4"),
        )

        with self.assertRaises(ProductMonthClosureError):
            self.service.build(month="2025-09")

    @override_settings(PRODUCT_MONTH_CLOSURE_SALES_SOURCE_MODE="OFFICIAL_MONTHLY_REPORT")
    def test_build_uses_official_monthly_report_for_sales_when_configured(self):
        previous = ProductoMonthClosure.objects.create(
            month_start=date(2025, 8, 1),
            month_end=date(2025, 8, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_BOOTSTRAP_SEED,
            opening_reference_date=date(2025, 8, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=self.parent,
            inventario_inicial_teorico=Decimal("10"),
            inventario_final_teorico=Decimal("10"),
        )
        PointProductionLine.objects.create(
            branch=self.point_branch,
            erp_branch=self.sucursal,
            receta=self.parent,
            production_external_id="prod-official-1",
            detail_external_id="detail-official-1",
            source_hash="prod-official-hash-1",
            production_date=date(2025, 9, 5),
            item_name=self.parent.nombre,
            item_code=self.parent.codigo_point,
            produced_quantity=Decimal("15"),
        )
        VentaHistorica.objects.create(
            receta=self.parent,
            sucursal=self.sucursal,
            fecha=date(2025, 9, 10),
            cantidad=Decimal("999"),
            fuente="POINT_BRIDGE_SALES",
        )

        parent_code = self.parent.codigo_point
        parent_name = self.parent.nombre

        class FakeOfficialSalesReportService:
            def fetch_report(self, **kwargs):
                return type(
                    "Report",
                    (),
                    {
                        "report_path": "/tmp/official-september.xls",
                        "request_url": "https://point.example/report",
                    },
                )()

            def parse_report(self, *, report_path: str):
                return type(
                    "ParsedReport",
                    (),
                    {
                        "rows": [
                            {
                                "Codigo": parent_code,
                                "Nombre": parent_name,
                                "Cantidad": Decimal("7"),
                            }
                        ],
                        "summary": {"venta": Decimal("100")},
                    },
                )()

        self.service.official_sales_report_service = FakeOfficialSalesReportService()

        closure = self.service.build(month="2025-09")

        line = closure.lines.get(receta_padre=self.parent)
        self.assertEqual(line.venta_directa_enteros, Decimal("999"))
        sales_meta = (closure.metadata or {}).get("sales_meta", {})
        self.assertEqual(sales_meta.get("configured_source_mode"), "OFFICIAL_MONTHLY_REPORT")
        self.assertFalse(sales_meta.get("authoritative"))

    @override_settings(PRODUCT_MONTH_CLOSURE_SALES_SOURCE_MODE="AUTO")
    def test_build_falls_back_to_official_point_daily_sales_when_monthly_report_fails(self):
        previous = ProductoMonthClosure.objects.create(
            month_start=date(2025, 8, 1),
            month_end=date(2025, 8, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_BOOTSTRAP_SEED,
            opening_reference_date=date(2025, 8, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=self.parent,
            inventario_inicial_teorico=Decimal("10"),
            inventario_final_teorico=Decimal("10"),
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=PointProduct.objects.create(
                external_id="parent-fallback-1",
                sku=self.parent.codigo_point,
                name=self.parent.nombre,
                category="Pasteles",
            ),
            receta=self.parent,
            sync_job=self.sync_job,
            sale_date=date(2025, 9, 10),
            quantity=Decimal("9"),
            tickets=0,
            gross_amount=Decimal("900"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("900"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("900"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )
        VentaHistorica.objects.create(
            receta=self.parent,
            sucursal=self.sucursal,
            fecha=date(2025, 9, 10),
            cantidad=Decimal("999"),
            fuente="POINT_BRIDGE_SALES",
        )

        class FailingOfficialSalesReportService:
            def fetch_report(self, **kwargs):
                raise RuntimeError("Point 500")

        self.service.official_sales_report_service = FailingOfficialSalesReportService()

        closure = self.service.build(month="2025-09")

        line = closure.lines.get(receta_padre=self.parent)
        self.assertEqual(line.venta_directa_enteros, Decimal("9"))
        sales_meta = (closure.metadata or {}).get("sales_meta", {})
        self.assertEqual(sales_meta.get("selected_source"), "official_point_daily_sales")

    @override_settings(PRODUCT_MONTH_CLOSURE_SALES_SOURCE_MODE="AUTO")
    def test_lock_rejects_closure_when_official_daily_sales_job_is_partial(self):
        previous = ProductoMonthClosure.objects.create(
            month_start=date(2025, 8, 1),
            month_end=date(2025, 8, 31),
            status=ProductoMonthClosure.STATUS_BUILT,
            opening_source=ProductoMonthClosure.OPENING_SOURCE_BOOTSTRAP_SEED,
            opening_reference_date=date(2025, 8, 31),
        )
        ProductoMonthClosureLine.objects.create(
            closure=previous,
            receta_padre=self.parent,
            inventario_inicial_teorico=Decimal("10"),
            inventario_final_teorico=Decimal("10"),
        )
        partial_job = PointSyncJob.objects.create(
            job_type=PointSyncJob.JOB_TYPE_SALES,
            status=PointSyncJob.STATUS_PARTIAL,
            parameters={
                "source": "POINT_OFFICIAL_REPORT",
                "start_date": "2025-09-01",
                "end_date": "2025-09-30",
                "branch_filter": "",
                "credito_scopes": ["null"],
                "excluded_ranges": [],
                "max_days": None,
            },
            error_message="Backfill oficial completado con 1 branch-day(s) omitidos por error.",
        )
        PointDailySale.objects.create(
            branch=self.point_branch,
            product=PointProduct.objects.create(
                external_id="parent-partial-1",
                sku=self.parent.codigo_point,
                name=self.parent.nombre,
                category="Pasteles",
            ),
            receta=self.parent,
            sync_job=partial_job,
            sale_date=date(2025, 9, 10),
            quantity=Decimal("9"),
            tickets=0,
            gross_amount=Decimal("900"),
            discount_amount=Decimal("0"),
            total_amount=Decimal("900"),
            tax_amount=Decimal("0"),
            net_amount=Decimal("900"),
            source_endpoint="/Report/PrintReportes?idreporte=3",
        )

        class FailingOfficialSalesReportService:
            def fetch_report(self, **kwargs):
                raise RuntimeError("Point 500")

        self.service.official_sales_report_service = FailingOfficialSalesReportService()

        closure = self.service.build(month="2025-09")

        validation = dict((closure.metadata or {}).get("validation") or {})
        sales_meta = dict((closure.metadata or {}).get("sales_meta") or {})
        self.assertFalse(validation.get("lock_ready"))
        self.assertIn("MONTH_SOURCE_INCOMPLETE", validation.get("blocking_issues") or [])
        self.assertIn("SALES_SYNC_JOB_PARTIAL", sales_meta.get("authority_issues") or [])
        with self.assertRaises(ProductMonthClosureError):
            self.service.lock(closure=closure, reason="test")
