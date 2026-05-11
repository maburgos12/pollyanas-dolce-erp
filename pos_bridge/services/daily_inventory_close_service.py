from __future__ import annotations

import unicodedata
from datetime import date, datetime, time, timedelta, timezone as datetime_timezone
from decimal import Decimal
from zoneinfo import ZoneInfo

from django.conf import settings
from django.db import connection
from django.db.models import Min, Max
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.models import Sucursal, sucursales_operativas
from pos_bridge.models import PointBranch, PointInventorySnapshot


ZERO = Decimal("0.000")
DAILY_CLOSE_CATEGORY_ORDER = [
    "Bollo",
    "Pastel Mini",
    "Pastel Chico",
    "Pastel Mediano",
    "Pastel Grande",
    "Pay Mediano",
    "Pay Grande",
    "Pastel Rebanada",
    "Pay Rebanada",
]


def _normalize(value: str | None) -> str:
    return " ".join((value or "").strip().lower().replace("_", " ").split())


def _branch_label(branch: Sucursal) -> str:
    return (branch.nombre or branch.codigo or "").replace("_", " ").strip() or branch.codigo


def _report_category(row: dict) -> str:
    category = (row.get("category") or "").strip()
    product_name = _normalize(row.get("product_name"))
    category_norm = _normalize(category)
    if category_norm == "rebanada":
        return "Pay Rebanada" if "pay" in product_name else "Pastel Rebanada"
    if category_norm == "bollo":
        return "Bollo"
    return category


class DailyInventoryCloseService:
    """Builds the end-of-day Point inventory matrix without duplicating snapshots."""

    def __init__(self, *, timezone_name: str | None = None):
        self.timezone_name = timezone_name or getattr(settings, "TIME_ZONE", "America/Mazatlan")
        self.local_tz = ZoneInfo(self.timezone_name)

    def _day_bounds_utc(self, fecha_operacion: date) -> tuple[datetime, datetime]:
        start_local = datetime.combine(fecha_operacion, time.min, tzinfo=self.local_tz)
        end_local = start_local + timedelta(days=1)
        return start_local.astimezone(datetime_timezone.utc), end_local.astimezone(datetime_timezone.utc)

    def _target_branches(self, fecha_operacion: date) -> list[Sucursal]:
        branches = list(sucursales_operativas(fecha_operacion).order_by("codigo", "nombre"))
        cedis = Sucursal.objects.filter(codigo="CEDIS").first()
        if cedis and all(branch.pk != cedis.pk for branch in branches):
            branches.append(cedis)
        return branches

    def build_close(self, *, fecha_operacion: date, category_filter: list[str] | None = None) -> dict:
        target_branches = self._target_branches(fecha_operacion)
        target_codes = [branch.codigo for branch in target_branches]
        point_branches = list(
            PointBranch.objects.filter(erp_branch_id__in=[branch.id for branch in target_branches]).select_related(
                "erp_branch"
            )
        )
        start_utc, end_utc = self._day_bounds_utc(fecha_operacion)
        snapshots = PointInventorySnapshot.objects.filter(
            branch_id__in=[branch.id for branch in point_branches],
            captured_at__gte=start_utc,
            captured_at__lt=end_utc,
        ).select_related("branch", "branch__erp_branch", "product")
        if connection.vendor == "postgresql":
            snapshots = snapshots.order_by("branch__erp_branch_id", "product_id", "-captured_at", "-id").distinct(
                "branch__erp_branch_id",
                "product_id",
            )
        else:
            snapshots = snapshots.order_by("branch__erp_branch__codigo", "product__name", "captured_at", "id")

        latest_by_branch_product: dict[tuple[str, int], PointInventorySnapshot] = {}
        branch_codes_with_capture: set[str] = set()
        for snapshot in snapshots:
            if not snapshot.branch.erp_branch_id:
                continue
            code = snapshot.branch.erp_branch.codigo
            if code not in target_codes:
                continue
            latest_by_branch_product[(code, snapshot.product_id)] = snapshot
            branch_codes_with_capture.add(code)

        product_map: dict[int, dict] = {}
        for (branch_code, product_id), snapshot in latest_by_branch_product.items():
            row = product_map.setdefault(
                product_id,
                {
                    "product_id": product_id,
                    "sku": snapshot.product.sku or snapshot.product.external_id,
                    "product_name": snapshot.product.name,
                    "category": snapshot.product.category,
                    "stocks": {code: ZERO for code in target_codes},
                    "captured_at_by_branch": {},
                    "total_stock": ZERO,
                },
            )
            stock = Decimal(str(snapshot.stock or 0)).quantize(ZERO)
            row["stocks"][branch_code] = stock
            row["captured_at_by_branch"][branch_code] = timezone.localtime(snapshot.captured_at, self.local_tz)

        category_filter_set = {_normalize(category) for category in category_filter or []}
        order_lookup = {_normalize(category): idx for idx, category in enumerate(DAILY_CLOSE_CATEGORY_ORDER)}
        rows = []
        for row in product_map.values():
            report_category = _report_category(row)
            row["report_category"] = report_category
            if category_filter_set and _normalize(report_category) not in category_filter_set:
                continue
            rows.append(row)

        if category_filter_set:
            rows = sorted(
                rows,
                key=lambda item: (
                    order_lookup.get(_normalize(item.get("report_category")), 999),
                    item.get("product_name") or "",
                    item.get("sku") or "",
                ),
            )
        else:
            rows = sorted(rows, key=lambda item: (item["product_name"] or "", item["sku"] or ""))
        for row in rows:
            row["total_stock"] = sum((row["stocks"].get(code, ZERO) for code in target_codes), ZERO).quantize(ZERO)

        capture_range = PointInventorySnapshot.objects.filter(
            id__in=[snapshot.id for snapshot in latest_by_branch_product.values()]
        ).aggregate(first=Min("captured_at"), last=Max("captured_at"))
        first_capture = capture_range["first"]
        last_capture = capture_range["last"]
        return {
            "fecha_operacion": fecha_operacion,
            "timezone_name": self.timezone_name,
            "branches": [{"code": branch.codigo, "name": branch.nombre, "label": _branch_label(branch)} for branch in target_branches],
            "rows": rows,
            "category_filter": category_filter or [],
            "missing_branch_codes": [code for code in target_codes if code not in branch_codes_with_capture],
            "first_capture_at": timezone.localtime(first_capture, self.local_tz) if first_capture else None,
            "last_capture_at": timezone.localtime(last_capture, self.local_tz) if last_capture else None,
            "generated_at": timezone.localtime(timezone.now(), self.local_tz),
        }

    def build_workbook(self, payload: dict) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario final"
        branch_codes = [branch["code"] for branch in payload["branches"]]
        max_col = 3 + len(branch_codes) + 1

        ws.cell(row=1, column=1, value="Inventario final al cierre")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.cell(row=2, column=1, value="Fecha operativa")
        ws.cell(row=2, column=2, value=payload["fecha_operacion"].isoformat())
        ws.cell(row=3, column=1, value="Zona horaria")
        ws.cell(row=3, column=2, value=payload["timezone_name"])
        ws.cell(row=4, column=1, value="Ultima captura Point")
        ws.cell(row=4, column=2, value=payload["last_capture_at"].strftime("%Y-%m-%d %H:%M") if payload["last_capture_at"] else "Sin captura")

        branch_labels = [branch.get("label") or branch["code"].replace("_", " ") for branch in payload["branches"]]
        header = ["SKU", "Producto", "Categoria", *branch_labels, "Total cierre"]
        ws.append(header)
        header_row = 5
        fill = PatternFill("solid", fgColor="1F2937")
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

        for row in payload["rows"]:
            ws.append(
                [
                    row["sku"],
                    row["product_name"],
                    row.get("report_category") or row["category"],
                    *[float(row["stocks"].get(code, ZERO)) for code in branch_codes],
                    float(row["total_stock"]),
                ]
            )

        widths = {"A": 14, "B": 42, "C": 18}
        for idx, code in enumerate(branch_codes, start=4):
            widths[get_column_letter(idx)] = max(12, min(18, len(code) + 4))
        widths[get_column_letter(max_col)] = 14
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
        for row in ws.iter_rows(min_row=6, min_col=4, max_col=max_col):
            for cell in row:
                cell.number_format = "#,##0.000"
                cell.alignment = Alignment(horizontal="right")
        ws.freeze_panes = "D6"
        return wb

    def build_pdf_bytes(self, payload: dict) -> bytes:
        def _pdf_safe(text: str) -> str:
            normalized = unicodedata.normalize("NFKD", str(text or ""))
            return normalized.encode("ascii", errors="ignore").decode("ascii")

        def _escape(text: str) -> str:
            return _pdf_safe(text).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

        def _label(branch: dict) -> str:
            return str(branch.get("label") or branch.get("code") or "").replace("_", " ")

        def _short(text: str, width: int) -> str:
            value = " ".join(str(text or "").replace("\n", " ").split())
            if len(value) <= width:
                return value.ljust(width)
            return value[:width]

        def _num(value: Decimal, width: int = 8) -> str:
            return f"{value:,.1f}".replace(",", "")[-width:].rjust(width)

        def _text_line(x: int, y: int, text: str, *, font: str = "F1", size: int = 7) -> str:
            return f"BT /{font} {size} Tf 1 0 0 1 {x} {y} Tm ({_escape(text)}) Tj ET"

        branch_totals = {branch["code"]: ZERO for branch in payload["branches"]}
        for row in payload["rows"]:
            for code, stock in row["stocks"].items():
                branch_totals[code] = (branch_totals.get(code, ZERO) + stock).quantize(ZERO)
        total = sum(branch_totals.values(), ZERO).quantize(ZERO)
        last_capture = payload["last_capture_at"].strftime("%Y-%m-%d %H:%M") if payload["last_capture_at"] else "Sin captura"
        branch_codes = [branch["code"] for branch in payload["branches"]]
        branch_headers = [_short(_label(branch).upper(), 7) for branch in payload["branches"]]
        header = (
            f"{'SKU':<10} {'PRODUCTO':<30} {'CATEGORIA':<16} {'TOTAL':>8} "
            + " ".join(branch_headers)
        )
        separator = "-" * len(header)

        summary_lines = [
            f"Fecha operativa: {payload['fecha_operacion'].isoformat()}",
            f"Zona horaria: {payload['timezone_name']}",
            f"Ultima captura Point: {last_capture}",
            f"Productos con inventario: {len(payload['rows'])}",
            f"Total inventario cierre: {total}",
        ]
        branch_total_parts = [
            f"{_label(branch)} {_num(branch_totals[branch['code']], 7).strip()}" for branch in payload["branches"]
        ]
        summary_lines.append("Totales por sucursal: " + " | ".join(branch_total_parts[:5]))
        summary_lines.append("Sucursales cont.: " + " | ".join(branch_total_parts[5:]))
        if payload["missing_branch_codes"]:
            summary_lines.append("Sin captura: " + ", ".join(payload["missing_branch_codes"]))

        row_lines = []
        for row in payload["rows"]:
            row_lines.append(
                f"{_short(row.get('sku'), 10)} "
                f"{_short(row.get('product_name'), 30)} "
                f"{_short(row.get('report_category') or row.get('category'), 16)} "
                f"{_num(row.get('total_stock') or ZERO)} "
                + " ".join(_num(row["stocks"].get(code, ZERO), 7) for code in branch_codes)
            )

        first_page_rows = 38
        next_page_rows = 45
        chunks = [row_lines[:first_page_rows]]
        remaining = row_lines[first_page_rows:]
        while remaining:
            chunks.append(remaining[:next_page_rows])
            remaining = remaining[next_page_rows:]

        page_contents: list[bytes] = []
        total_pages = max(1, len(chunks))
        for page_index, chunk in enumerate(chunks or [[]], start=1):
            y = 570
            lines = [
                _text_line(30, y, "Inventario final al cierre", font="F2", size=14),
                _text_line(700, y, f"Pagina {page_index}/{total_pages}", font="F2", size=8),
            ]
            y -= 18
            if page_index == 1:
                for summary in summary_lines:
                    lines.append(_text_line(30, y, summary, font="F2", size=8))
                    y -= 12
                y -= 4
            lines.append(_text_line(30, y, header, font="F1", size=6))
            y -= 9
            lines.append(_text_line(30, y, separator, font="F1", size=6))
            y -= 9
            for row_line in chunk:
                lines.append(_text_line(30, y, row_line, font="F1", size=6))
                y -= 10
            page_contents.append("\n".join(lines).encode("latin-1", errors="replace"))

        objects: list[bytes] = [
            b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj",
            b"3 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Courier >> endobj",
            b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >> endobj",
        ]
        page_ids = []
        content_ids = []
        next_obj_id = 5
        for content in page_contents:
            page_id = next_obj_id
            content_id = next_obj_id + 1
            next_obj_id += 2
            page_ids.append(page_id)
            content_ids.append(content_id)
            objects.append(
                f"{page_id} 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 792 612] "
                f"/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents {content_id} 0 R >> endobj".encode()
            )
            objects.append(
                b"%d 0 obj << /Length " % content_id
                + str(len(content)).encode()
                + b" >> stream\n"
                + content
                + b"\nendstream endobj"
            )
        objects.insert(1, f"2 0 obj << /Type /Pages /Count {len(page_ids)} /Kids ".encode() + b"[" + b" ".join(f"{page_id} 0 R".encode() for page_id in page_ids) + b"] >> endobj")

        output = bytearray(b"%PDF-1.4\n")
        offsets = [0]
        for obj in objects:
            offsets.append(len(output))
            output.extend(obj)
            output.extend(b"\n")
        xref_pos = len(output)
        size = max(4, *(page_ids or [0]), *(content_ids or [0])) + 1
        offset_by_id = {0: 0}
        for offset, obj in zip(offsets[1:], objects):
            obj_id = int(obj.split(b" ", 1)[0])
            offset_by_id[obj_id] = offset
        output.extend(f"xref\n0 {size}\n".encode())
        output.extend(b"0000000000 65535 f \n")
        for obj_id in range(1, size):
            offset = offset_by_id.get(obj_id, 0)
            output.extend(f"{offset:010d} 00000 n \n".encode())
        output.extend(f"trailer << /Size {size} /Root 1 0 R >>\nstartxref\n{xref_pos}\n%%EOF".encode())
        return bytes(output)
