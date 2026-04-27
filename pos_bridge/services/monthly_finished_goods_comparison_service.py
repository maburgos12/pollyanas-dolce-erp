from __future__ import annotations

from calendar import monthrange
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
import json
from pathlib import Path
from typing import Any

from django.db.models import Max, Sum
from django.utils import timezone
from openpyxl import Workbook

from pos_bridge.models import PointProductionLine, PointWasteLine
from pos_bridge.services.movement_sync_service import PointMovementSyncService
from pos_bridge.services.official_sales_backfill_service import OfficialSalesBackfillService
from pos_bridge.services.product_month_closure_service import ProductMonthClosureService
from recetas.models import ProductoMonthClosure, Receta, RecetaEquivalencia, RecetaPresentacionDerivada, VentaHistorica
from recetas.utils.cierre_equivalencias import resolve_closure_recipe_quantity
from ventas.services.sales_canonical_source import POINT_BRIDGE_SALES_SOURCE


@dataclass(frozen=True)
class CoverageSnapshot:
    sales_rows: int
    production_rows: int
    waste_rows: int
    sales_latest_date: str
    production_latest_date: str
    waste_latest_date: str

    def as_dict(self) -> dict[str, Any]:
        return {
            "sales_rows": self.sales_rows,
            "production_rows": self.production_rows,
            "waste_rows": self.waste_rows,
            "sales_latest_date": self.sales_latest_date,
            "production_latest_date": self.production_latest_date,
            "waste_latest_date": self.waste_latest_date,
        }


class MonthlyFinishedGoodsComparisonService:
    DEFAULT_OUTPUT_DIR = Path("output/monthly_finished_goods_comparison")

    def __init__(
        self,
        *,
        official_sales_backfill_service: OfficialSalesBackfillService | None = None,
        movement_sync_service: PointMovementSyncService | None = None,
        closure_service: ProductMonthClosureService | None = None,
    ):
        self.official_sales_backfill_service = official_sales_backfill_service or OfficialSalesBackfillService()
        self.movement_sync_service = movement_sync_service or PointMovementSyncService()
        self.closure_service = closure_service or ProductMonthClosureService()

    def run(
        self,
        *,
        month: str | date,
        output_dir: str | Path | None = None,
        triggered_by=None,
        rebuild: bool = False,
        fallback_to_point: bool = True,
        branch_filter: str | None = None,
    ) -> dict[str, Any]:
        month_start = self._parse_month(month)
        month_end = self._month_end(month_start)

        coverage_before = self.inspect_month_coverage(month_start=month_start, month_end=month_end)
        sync_actions: list[dict[str, Any]] = []

        if fallback_to_point:
            sync_actions.extend(
                self._ensure_month_data(
                    month_start=month_start,
                    month_end=month_end,
                    branch_filter=branch_filter,
                    triggered_by=triggered_by,
                    coverage=coverage_before,
                )
            )

        coverage_after = self.inspect_month_coverage(month_start=month_start, month_end=month_end)
        effective_rebuild = rebuild or bool(sync_actions)
        closure = self._resolve_or_build_closure(
            month_start=month_start,
            rebuild=effective_rebuild,
            triggered_by=triggered_by,
        )
        export_path = self.export_closure_xlsx(
            closure=closure,
            output_dir=output_dir or self.DEFAULT_OUTPUT_DIR,
            coverage_before=coverage_before,
            coverage_after=coverage_after,
            sync_actions=sync_actions,
        )

        return {
            "month": month_start.strftime("%Y-%m"),
            "month_start": month_start.isoformat(),
            "month_end": month_end.isoformat(),
            "fallback_to_point": bool(fallback_to_point),
            "rebuild_requested": bool(rebuild),
            "rebuild_applied": bool(effective_rebuild),
            "coverage_before": coverage_before.as_dict(),
            "coverage_after": coverage_after.as_dict(),
            "sync_actions": sync_actions,
            "closure_id": closure.id,
            "closure_status": closure.status,
            "closure_locked": closure.is_locked,
            "export_path": str(export_path),
        }

    def dry_run(self, *, month: str | date) -> dict[str, Any]:
        month_start = self._parse_month(month)
        month_end = self._month_end(month_start)
        coverage = self.inspect_month_coverage(month_start=month_start, month_end=month_end)
        operational_validation = self.validate_operational_readiness(month=month_start)
        payload: dict[str, Any] = {
            "month": month_start.strftime("%Y-%m"),
            "month_start": month_start.isoformat(),
            "month_end": month_end.isoformat(),
            "dry_run": True,
            "persisted": False,
            "coverage": coverage.as_dict(),
            "operational_validation": operational_validation,
        }
        try:
            plan = self.closure_service.preview(month=month_start)
        except Exception as exc:  # noqa: BLE001
            payload["closure_preview_error"] = str(exc)
            return payload

        validation = dict((plan.get("metadata") or {}).get("validation") or {})
        payload["closure_preview"] = {
            "opening_source": plan.get("opening_source"),
            "opening_reference_date": plan.get("opening_reference_date"),
            "line_count": len(plan.get("line_rows") or []),
            "totals": {key: str(value) for key, value in dict(plan.get("totals") or {}).items()},
            "lock_ready": bool(validation.get("lock_ready")),
            "warnings": list(validation.get("warnings") or []),
            "blocking_issues": list(validation.get("blocking_issues") or []),
        }
        return payload

    def validate_operational_readiness(self, *, month: str | date) -> dict[str, Any]:
        month_start = self._parse_month(month)
        month_end = self._month_end(month_start)
        production_buckets, production_raw = self._production_finished_good_buckets(month_start=month_start, month_end=month_end)
        sales_buckets, sales_raw = self._sales_finished_good_buckets(month_start=month_start, month_end=month_end)

        production_ids = set(production_buckets)
        sales_ids = set(sales_buckets)
        overlap_ids = production_ids & sales_ids
        sales_only_ids = sales_ids - production_ids
        production_only_ids = production_ids - sales_ids

        total_produced = sum((bucket["units"] for bucket in production_buckets.values()), Decimal("0"))
        total_sold = sum((bucket["units"] for bucket in sales_buckets.values()), Decimal("0"))
        global_diff_pct = self._pct((total_produced - total_sold), total_sold)
        product_coverage_pct = self._pct(Decimal(len(overlap_ids)), Decimal(len(sales_ids))) if sales_ids else Decimal("0")

        over_3x_rows = []
        within_30_count = 0
        compared_count = 0
        for receta_id in sorted(overlap_ids):
            produced = production_buckets[receta_id]["units"]
            sold = sales_buckets[receta_id]["units"]
            if sold > 0:
                compared_count += 1
                diff_pct = self._pct(produced - sold, sold)
                if abs(diff_pct) <= Decimal("30"):
                    within_30_count += 1
                if produced <= 0 or sold > produced * Decimal("3"):
                    over_3x_rows.append(self._comparison_payload(receta_id, production_buckets, sales_buckets, diff_pct=diff_pct))

        criteria = {
            "coverage_pct_min_70": product_coverage_pct >= Decimal("70"),
            "global_diff_abs_pct_max_40": abs(global_diff_pct) <= Decimal("40"),
            "no_sales_gt_3x_production": not over_3x_rows,
        }
        sales_only_payload = self._sales_only_payload(sales_only_ids=sales_only_ids, sales_buckets=sales_buckets)[:35]

        return {
            "month": month_start.strftime("%Y-%m"),
            "passed": all(criteria.values()),
            "criteria": criteria,
            "production_rows_after_filter": int(production_raw["rows_after_filter"]),
            "production_rows_before_filter": int(production_raw["rows_before_filter"]),
            "production_recipes_after_filter": len(production_ids),
            "sales_recipes": len(sales_ids),
            "recipes_in_both": len(overlap_ids),
            "recipes_only_sales": len(sales_only_ids),
            "recipes_only_production": len(production_only_ids),
            "total_produced_units": str(total_produced),
            "total_sold_units": str(total_sold),
            "global_diff_pct": str(global_diff_pct),
            "sales_product_coverage_pct": str(product_coverage_pct),
            "products_within_30_pct_count": within_30_count,
            "products_compared_count": compared_count,
            "sales_gt_3x_production_count": len(over_3x_rows),
            "sales_gt_3x_production": over_3x_rows[:10],
            "top_production": self._top_bucket_payload(production_buckets)[:10],
            "top_sales": self._top_bucket_payload(sales_buckets)[:10],
            "sales_only_with_parent": sales_only_payload,
            "raw_sales_only_with_parent": self._raw_sales_only_with_parent(month_start=month_start, month_end=month_end)[:35],
        }

    def inspect_month_coverage(self, *, month_start: date, month_end: date) -> CoverageSnapshot:
        sales_qs = VentaHistorica.objects.filter(
            fecha__gte=month_start,
            fecha__lte=month_end,
            fuente=POINT_BRIDGE_SALES_SOURCE,
        )
        production_qs = PointProductionLine.objects.filter(
            production_date__gte=month_start,
            production_date__lte=month_end,
            receta__isnull=False,
            receta__tipo=Receta.TIPO_PRODUCTO_FINAL,
        )
        production_qs = production_qs.exclude(receta__modo_costeo=Receta.MODO_COSTEO_SERVICIO)
        waste_qs = PointWasteLine.objects.filter(
            movement_at__date__gte=month_start,
            movement_at__date__lte=month_end,
            receta__isnull=False,
        )

        return CoverageSnapshot(
            sales_rows=sales_qs.count(),
            production_rows=production_qs.count(),
            waste_rows=waste_qs.count(),
            sales_latest_date=self._date_to_text(sales_qs.aggregate(value=Max("fecha"))["value"]),
            production_latest_date=self._date_to_text(production_qs.aggregate(value=Max("production_date"))["value"]),
            waste_latest_date=self._date_to_text(waste_qs.aggregate(value=Max("movement_at"))["value"]),
        )

    def export_closure_xlsx(
        self,
        *,
        closure,
        output_dir: str | Path,
        coverage_before: CoverageSnapshot,
        coverage_after: CoverageSnapshot,
        sync_actions: list[dict[str, Any]],
    ) -> Path:
        output_root = Path(output_dir)
        output_root.mkdir(parents=True, exist_ok=True)
        output_path = output_root / f"comparativa_producto_terminado_{closure.month_start:%Y-%m}.xlsx"
        direct_payload = self._build_direct_monthly_payload(month_start=closure.month_start, month_end=closure.month_end)

        wb = Workbook()
        summary_ws = wb.active
        summary_ws.title = "RESUMEN_DIRECTO"
        self._write_direct_summary_sheet(
            ws=summary_ws,
            month_start=closure.month_start,
            month_end=closure.month_end,
            direct_payload=direct_payload,
            coverage_before=coverage_before,
            coverage_after=coverage_after,
            sync_actions=sync_actions,
        )

        direct_ws = wb.create_sheet("COMPARATIVA_DIRECTA")
        self._write_direct_detail_sheet(ws=direct_ws, rows=direct_payload["comparison_rows"])

        production_raw_ws = wb.create_sheet("PRODUCCION_RAW")
        self._write_raw_sheet(ws=production_raw_ws, rows=direct_payload["production_raw"])

        sales_raw_ws = wb.create_sheet("VENTAS_RAW")
        self._write_raw_sheet(ws=sales_raw_ws, rows=direct_payload["sales_raw"])

        waste_raw_ws = wb.create_sheet("MERMAS_RAW")
        self._write_raw_sheet(ws=waste_raw_ws, rows=direct_payload["waste_raw"])

        closure_ws = wb.create_sheet("CIERRE_TEORICO")
        self._write_closure_summary_sheet(
            ws=closure_ws,
            closure=closure,
        )

        closure_detail_ws = wb.create_sheet("CIERRE_TEO_DETALLE")
        self._write_closure_detail_sheet(ws=closure_detail_ws, closure=closure)

        coverage_ws = wb.create_sheet("COBERTURA")
        self._write_coverage_sheet(
            ws=coverage_ws,
            coverage_before=coverage_before,
            coverage_after=coverage_after,
        )

        jobs_ws = wb.create_sheet("SYNC_ACTIONS")
        self._write_sync_actions_sheet(ws=jobs_ws, sync_actions=sync_actions)

        wb.save(output_path)
        return output_path

    def _build_direct_monthly_payload(self, *, month_start: date, month_end: date) -> dict[str, Any]:
        buckets: dict[str, dict[str, Any]] = {}
        production_raw: list[dict[str, Any]] = []
        sales_raw: list[dict[str, Any]] = []
        waste_raw: list[dict[str, Any]] = []

        for movement in (
            PointProductionLine.objects.select_related("receta")
            .filter(
                production_date__gte=month_start,
                production_date__lte=month_end,
                receta__isnull=False,
                receta__tipo=Receta.TIPO_PRODUCTO_FINAL,
            )
            .exclude(receta__excluir_cierre=True)
            .exclude(receta__modo_costeo=Receta.MODO_COSTEO_SERVICIO)
            .order_by("production_date", "id")
        ):
            receta = movement.receta
            if not self.closure_service._is_recipe_eligible_for_closure(receta):
                continue
            code = (receta.codigo_point or "").strip()
            if not code:
                continue
            bucket = self._comparison_bucket(buckets, code=code, product_name=receta.nombre)
            quantity = float(movement.produced_quantity or 0)
            bucket["ingreso_produccion"] += quantity
            bucket["production_refs"].add((movement.production_external_id or "").strip())
            raw_row = {
                "fecha": self._date_to_text(movement.production_date),
                "codigo_point": code,
                "producto": receta.nombre,
                "cantidad": quantity,
                "referencia": movement.production_external_id or "",
                "tipo": "POINT_PRODUCTION_LINE",
            }
            production_raw.append(raw_row)

        for sale in (
            VentaHistorica.objects.select_related("receta", "sucursal")
            .filter(
                fecha__gte=month_start,
                fecha__lte=month_end,
                fuente=POINT_BRIDGE_SALES_SOURCE,
                receta__isnull=False,
            )
            .exclude(receta__excluir_cierre=True)
            .order_by("fecha", "id")
        ):
            receta = sale.receta
            if not self.closure_service._is_recipe_eligible_for_closure(receta):
                continue
            code = (receta.codigo_point or "").strip()
            if not code:
                continue
            bucket = self._comparison_bucket(buckets, code=code, product_name=receta.nombre)
            quantity = float(sale.cantidad or 0)
            bucket["venta_total_sucursales"] += quantity
            if sale.sucursal_id:
                bucket["sales_branches"].add(sale.sucursal.codigo)
            raw_row = {
                "fecha": self._date_to_text(sale.fecha),
                "codigo_point": code,
                "producto": receta.nombre,
                "cantidad": quantity,
                "sucursal": sale.sucursal.codigo if sale.sucursal_id else "",
                "fuente": sale.fuente,
            }
            sales_raw.append(raw_row)

        for waste in (
            PointWasteLine.objects.select_related("receta", "erp_branch", "branch")
            .filter(
                movement_at__date__gte=month_start,
                movement_at__date__lte=month_end,
                receta__isnull=False,
            )
            .exclude(receta__excluir_cierre=True)
            .order_by("movement_at", "id")
        ):
            receta = waste.receta
            if not self.closure_service._is_recipe_eligible_for_closure(receta):
                continue
            code = (receta.codigo_point or waste.item_code or "").strip()
            if not code:
                continue
            bucket = self._comparison_bucket(buckets, code=code, product_name=receta.nombre or waste.item_name)
            quantity = float(waste.quantity or 0)
            bucket["merma_total_sucursales"] += quantity
            branch_code = ""
            if waste.erp_branch_id:
                branch_code = waste.erp_branch.codigo
            elif waste.branch_id:
                branch_code = waste.branch.external_id
            if branch_code:
                bucket["waste_branches"].add(branch_code)
            raw_row = {
                "fecha": self._date_to_text(waste.movement_at),
                "codigo_point": code,
                "producto": receta.nombre or waste.item_name,
                "cantidad": quantity,
                "sucursal": branch_code,
                "justificacion": waste.justification or "",
            }
            waste_raw.append(raw_row)

        comparison_rows: list[dict[str, Any]] = []
        for code in sorted(buckets):
            bucket = buckets[code]
            comparison_rows.append(
                {
                    "codigo_point": code,
                    "producto": bucket["producto"],
                    "ingreso_produccion": round(bucket["ingreso_produccion"], 3),
                    "venta_total_sucursales": round(bucket["venta_total_sucursales"], 3),
                    "merma_total_sucursales": round(bucket["merma_total_sucursales"], 3),
                    "saldo_directo": round(
                        bucket["ingreso_produccion"] - bucket["venta_total_sucursales"] - bucket["merma_total_sucursales"],
                        3,
                    ),
                    "sucursales_venta": ", ".join(sorted(bucket["sales_branches"])),
                    "sucursales_merma": ", ".join(sorted(bucket["waste_branches"])),
                    "referencias_produccion": ", ".join(sorted(ref for ref in bucket["production_refs"] if ref)),
                }
            )

        return {
            "comparison_rows": comparison_rows,
            "production_raw": production_raw,
            "sales_raw": sales_raw,
            "waste_raw": waste_raw,
        }

    def _ensure_month_data(
        self,
        *,
        month_start: date,
        month_end: date,
        branch_filter: str | None,
        triggered_by,
        coverage: CoverageSnapshot,
    ) -> list[dict[str, Any]]:
        actions: list[dict[str, Any]] = []

        if coverage.sales_rows <= 0:
            sales_job = self.official_sales_backfill_service.run(
                start_date=month_start,
                end_date=month_end,
                branch_filter=branch_filter,
                credito_scopes=["null"],
                triggered_by=triggered_by,
            )
            actions.append(
                self._build_sync_action(
                    kind="sales_official_backfill",
                    job=sales_job,
                    start_date=month_start,
                    end_date=month_end,
                )
            )

        if coverage.production_rows <= 0:
            production_job = self.movement_sync_service.run_production_sync(
                start_date=month_start,
                end_date=month_end,
                branch_filter=branch_filter,
                triggered_by=triggered_by,
            )
            actions.append(
                self._build_sync_action(
                    kind="production_sync",
                    job=production_job,
                    start_date=month_start,
                    end_date=month_end,
                )
            )

        if coverage.waste_rows <= 0:
            waste_job = self.movement_sync_service.run_waste_sync(
                start_date=month_start,
                end_date=month_end,
                branch_filter=branch_filter,
                triggered_by=triggered_by,
            )
            actions.append(
                self._build_sync_action(
                    kind="waste_sync",
                    job=waste_job,
                    start_date=month_start,
                    end_date=month_end,
                )
            )

        return actions

    def _production_finished_good_buckets(self, *, month_start: date, month_end: date) -> tuple[dict[int, dict[str, Any]], dict[str, Any]]:
        base_qs = PointProductionLine.objects.filter(
            production_date__gte=month_start,
            production_date__lte=month_end,
            receta__isnull=False,
        )
        rows = (
            base_qs.select_related("receta")
            .filter(receta__tipo=Receta.TIPO_PRODUCTO_FINAL)
            .exclude(receta__excluir_cierre=True)
            .exclude(receta__modo_costeo=Receta.MODO_COSTEO_SERVICIO)
            .order_by("id")
        )
        buckets: dict[int, dict[str, Any]] = {}
        for row in rows:
            parent_receta, qty = self._canonical_recipe_quantity(receta=row.receta, quantity=Decimal(str(row.produced_quantity or 0)))
            if parent_receta is None:
                continue
            bucket = self._recipe_bucket(buckets, parent_receta)
            bucket["units"] += qty
            bucket["rows"] += 1
            bucket["source_recipes"].add(row.receta.nombre)
        return buckets, {"rows_before_filter": base_qs.count(), "rows_after_filter": rows.count()}

    def _sales_finished_good_buckets(self, *, month_start: date, month_end: date) -> tuple[dict[int, dict[str, Any]], dict[str, Any]]:
        rows = (
            VentaHistorica.objects.select_related("receta")
            .filter(
                fecha__gte=month_start,
                fecha__lte=month_end,
                fuente=POINT_BRIDGE_SALES_SOURCE,
                receta__isnull=False,
                receta__tipo=Receta.TIPO_PRODUCTO_FINAL,
            )
            .exclude(receta__excluir_cierre=True)
            .exclude(receta__modo_costeo=Receta.MODO_COSTEO_SERVICIO)
            .order_by("id")
        )
        buckets: dict[int, dict[str, Any]] = {}
        for row in rows:
            parent_receta, qty = self._canonical_recipe_quantity(receta=row.receta, quantity=Decimal(str(row.cantidad or 0)))
            if parent_receta is None:
                continue
            bucket = self._recipe_bucket(buckets, parent_receta)
            bucket["units"] += qty
            bucket["rows"] += 1
            bucket["source_recipes"].add(row.receta.nombre)
        return buckets, {"rows": rows.count()}

    def _canonical_recipe_quantity(self, *, receta: Receta, quantity: Decimal) -> tuple[Receta | None, Decimal]:
        parent_receta, qty, _issue_note, _is_derived, _source = resolve_closure_recipe_quantity(receta, quantity)
        return parent_receta, qty

    def _recipe_bucket(self, buckets: dict[int, dict[str, Any]], receta: Receta) -> dict[str, Any]:
        bucket = buckets.get(receta.id)
        if bucket is None:
            bucket = {
                "receta_id": receta.id,
                "codigo_point": receta.codigo_point,
                "nombre": receta.nombre,
                "tipo": receta.tipo,
                "modo_costeo": receta.modo_costeo,
                "categoria": receta.categoria,
                "units": Decimal("0"),
                "rows": 0,
                "source_recipes": set(),
            }
            buckets[receta.id] = bucket
        return bucket

    def _top_bucket_payload(self, buckets: dict[int, dict[str, Any]]) -> list[dict[str, Any]]:
        rows = []
        for bucket in buckets.values():
            rows.append(
                {
                    "receta_id": bucket["receta_id"],
                    "codigo_point": bucket["codigo_point"],
                    "nombre": bucket["nombre"],
                    "categoria": bucket["categoria"],
                    "units": str(bucket["units"]),
                    "rows": bucket["rows"],
                    "source_recipes": sorted(bucket["source_recipes"])[:5],
                }
            )
        return sorted(rows, key=lambda row: Decimal(str(row["units"] or "0")), reverse=True)

    def _comparison_payload(
        self,
        receta_id: int,
        production_buckets: dict[int, dict[str, Any]],
        sales_buckets: dict[int, dict[str, Any]],
        *,
        diff_pct: Decimal,
    ) -> dict[str, Any]:
        sales_bucket = sales_buckets[receta_id]
        production_bucket = production_buckets.get(receta_id)
        produced = production_bucket["units"] if production_bucket else Decimal("0")
        return {
            "receta_id": receta_id,
            "codigo_point": sales_bucket["codigo_point"],
            "nombre": sales_bucket["nombre"],
            "produced_units": str(produced),
            "sold_units": str(sales_bucket["units"]),
            "diff_pct": str(diff_pct),
        }

    def _sales_only_payload(self, *, sales_only_ids: set[int], sales_buckets: dict[int, dict[str, Any]]) -> list[dict[str, Any]]:
        relation_by_parent = {
            row.receta_padre_id: row
            for row in RecetaPresentacionDerivada.objects.filter(receta_padre_id__in=sales_only_ids, activo=True)
            .select_related("receta_derivada", "receta_padre")
            .order_by("receta_padre_id", "id")
        }
        equivalence_by_parent = {
            row.receta_padre_id: row
            for row in RecetaEquivalencia.objects.filter(receta_padre_id__in=sales_only_ids, activo=True)
            .select_related("receta_porcion", "receta_padre")
            .order_by("receta_padre_id", "id")
        }
        payload = []
        for receta_id in sorted(sales_only_ids, key=lambda key: sales_buckets[key]["units"], reverse=True):
            bucket = sales_buckets[receta_id]
            relation = relation_by_parent.get(receta_id)
            equivalence = equivalence_by_parent.get(receta_id)
            payload.append(
                {
                    "receta_id": receta_id,
                    "codigo_point": bucket["codigo_point"],
                    "nombre": bucket["nombre"],
                    "categoria": bucket["categoria"],
                    "sold_units": str(bucket["units"]),
                    "possible_parent": bucket["nombre"],
                    "has_derived_children": relation is not None or equivalence is not None,
                    "example_derived_child": (
                        relation.receta_derivada.nombre
                        if relation
                        else equivalence.receta_porcion.nombre
                        if equivalence
                        else ""
                    ),
                    "units_per_parent": str(relation.unidades_por_padre) if relation else str(equivalence.factor_conversion) if equivalence else "",
                    "source_recipes": sorted(bucket["source_recipes"])[:8],
                }
            )
        return payload

    def _raw_sales_only_with_parent(self, *, month_start: date, month_end: date) -> list[dict[str, Any]]:
        production_recipe_ids = set(
            PointProductionLine.objects.filter(
                production_date__gte=month_start,
                production_date__lte=month_end,
                receta__isnull=False,
                receta__tipo=Receta.TIPO_PRODUCTO_FINAL,
            )
            .exclude(receta__excluir_cierre=True)
            .exclude(receta__modo_costeo=Receta.MODO_COSTEO_SERVICIO)
            .values_list("receta_id", flat=True)
            .distinct()
        )
        sales_rows = list(
            VentaHistorica.objects.filter(
                fecha__gte=month_start,
                fecha__lte=month_end,
                fuente=POINT_BRIDGE_SALES_SOURCE,
                receta__isnull=False,
                receta__tipo=Receta.TIPO_PRODUCTO_FINAL,
            )
            .exclude(receta__excluir_cierre=True)
            .exclude(receta__modo_costeo=Receta.MODO_COSTEO_SERVICIO)
            .exclude(receta_id__in=production_recipe_ids)
            .values(
                "receta_id",
                "receta__codigo_point",
                "receta__nombre",
                "receta__categoria",
                "receta__tipo",
                "receta__modo_costeo",
            )
            .annotate(units=Sum("cantidad"))
            .order_by("-units")
        )
        recipe_ids = [int(row["receta_id"]) for row in sales_rows]
        relations = {
            relation.receta_derivada_id: relation
            for relation in RecetaPresentacionDerivada.objects.filter(receta_derivada_id__in=recipe_ids, activo=True)
            .select_related("receta_padre", "receta_derivada")
            .order_by("receta_derivada_id", "id")
        }
        equivalences = {
            equivalence.receta_porcion_id: equivalence
            for equivalence in RecetaEquivalencia.objects.filter(receta_porcion_id__in=recipe_ids, activo=True)
            .select_related("receta_padre", "receta_porcion")
            .order_by("receta_porcion_id", "id")
        }
        payload = []
        for row in sales_rows:
            recipe_id = int(row["receta_id"])
            relation = relations.get(recipe_id)
            equivalence = equivalences.get(recipe_id)
            payload.append(
                {
                    "receta_id": recipe_id,
                    "codigo_point": row["receta__codigo_point"],
                    "nombre": row["receta__nombre"],
                    "categoria": row["receta__categoria"],
                    "sold_units": str(row["units"] or Decimal("0")),
                    "possible_parent": (
                        equivalence.receta_padre.nombre
                        if equivalence
                        else relation.receta_padre.nombre
                        if relation
                        else ""
                    ),
                    "parent_codigo_point": (
                        equivalence.receta_padre.codigo_point
                        if equivalence
                        else relation.receta_padre.codigo_point
                        if relation
                        else ""
                    ),
                    "units_per_parent": (
                        str(equivalence.factor_conversion)
                        if equivalence
                        else str(relation.unidades_por_padre)
                        if relation
                        else ""
                    ),
                    "has_formal_parent_relation": relation is not None or equivalence is not None,
                    "parent_relation_source": "RecetaEquivalencia" if equivalence else "RecetaPresentacionDerivada" if relation else "",
                }
            )
        return payload

    def _pct(self, numerator: Decimal, denominator: Decimal) -> Decimal:
        if denominator == 0:
            return Decimal("0")
        return (numerator / denominator * Decimal("100")).quantize(Decimal("0.01"))

    def _comparison_bucket(self, buckets: dict[str, dict[str, Any]], *, code: str, product_name: str) -> dict[str, Any]:
        bucket = buckets.get(code)
        if bucket is None:
            bucket = {
                "producto": product_name,
                "ingreso_produccion": 0.0,
                "venta_total_sucursales": 0.0,
                "merma_total_sucursales": 0.0,
                "sales_branches": set(),
                "waste_branches": set(),
                "production_refs": set(),
            }
            buckets[code] = bucket
        return bucket

    def _resolve_or_build_closure(self, *, month_start: date, rebuild: bool, triggered_by):
        existing = ProductoMonthClosure.objects.filter(month_start=month_start).order_by("-id").first()
        if existing is not None and not rebuild:
            return existing
        return self.closure_service.build(
            month=month_start,
            rebuild=rebuild,
            built_by=triggered_by,
            approval_reason="monthly_finished_goods_comparison",
            approval_channel="command_monthly_finished_goods_comparison",
        )

    def _write_direct_summary_sheet(
        self,
        *,
        ws,
        month_start: date,
        month_end: date,
        direct_payload: dict[str, Any],
        coverage_before: CoverageSnapshot,
        coverage_after: CoverageSnapshot,
        sync_actions,
    ):
        comparison_rows = list(direct_payload.get("comparison_rows") or [])
        total_production = sum((row["ingreso_produccion"] for row in comparison_rows), start=0.0)
        total_sales = sum((row["venta_total_sucursales"] for row in comparison_rows), start=0.0)
        total_waste = sum((row["merma_total_sucursales"] for row in comparison_rows), start=0.0)
        total_balance = sum((row["saldo_directo"] for row in comparison_rows), start=0.0)

        rows = [
            ["month", month_start.strftime("%Y-%m")],
            ["month_start", month_start.isoformat()],
            ["month_end", month_end.isoformat()],
            ["comparison_mode", "direct_raw_sources"],
            ["production_source", "MovimientoProductoCedis tipo=ENTRADA"],
            ["sales_source", "VentaHistorica fuente=POINT_BRIDGE_SALES"],
            ["waste_source", "PointWasteLine"],
            ["codigos_comparados", len(comparison_rows)],
            ["ingreso_produccion_total", round(total_production, 3)],
            ["venta_total_sucursales", round(total_sales, 3)],
            ["merma_total_sucursales", round(total_waste, 3)],
            ["saldo_directo", round(total_balance, 3)],
            ["coverage_before", json.dumps(coverage_before.as_dict(), ensure_ascii=False)],
            ["coverage_after", json.dumps(coverage_after.as_dict(), ensure_ascii=False)],
            ["sync_actions_count", len(sync_actions)],
        ]
        for row in rows:
            ws.append(row)

    def _write_closure_summary_sheet(self, *, ws, closure):
        lines = list(closure.lines.select_related("receta_padre").all())
        total_opening = sum((line.inventario_inicial_teorico for line in lines), start=0)
        total_production = sum((line.produccion_mes for line in lines), start=0)
        total_sales = sum((line.venta_total_equivalente for line in lines), start=0)
        total_waste = sum((line.merma_total_equivalente for line in lines), start=0)
        total_ending = sum((line.inventario_final_teorico for line in lines), start=0)

        rows = [
            ["month", closure.month_start.strftime("%Y-%m")],
            ["closure_status", closure.status],
            ["closure_locked", "YES" if closure.is_locked else "NO"],
            ["opening_source", closure.opening_source],
            ["opening_reference_date", self._date_to_text(closure.opening_reference_date)],
            ["built_at", self._date_to_text(closure.built_at)],
            ["built_by", getattr(closure.built_by, "username", "") if closure.built_by_id else ""],
            ["line_count", len(lines)],
            ["inventario_inicial_teorico", float(total_opening)],
            ["produccion_mes", float(total_production)],
            ["venta_total_equivalente", float(total_sales)],
            ["merma_total_equivalente", float(total_waste)],
            ["inventario_final_teorico", float(total_ending)],
            ["notes", closure.notes or ""],
        ]
        for row in rows:
            ws.append(row)

    def _write_direct_detail_sheet(self, *, ws, rows: list[dict[str, Any]]):
        ws.append(
            [
                "codigo_point",
                "producto",
                "ingreso_produccion",
                "venta_total_sucursales",
                "merma_total_sucursales",
                "saldo_directo",
                "sucursales_venta",
                "sucursales_merma",
                "referencias_produccion",
            ]
        )
        for row in rows:
            ws.append(
                [
                    row["codigo_point"],
                    row["producto"],
                    row["ingreso_produccion"],
                    row["venta_total_sucursales"],
                    row["merma_total_sucursales"],
                    row["saldo_directo"],
                    row["sucursales_venta"],
                    row["sucursales_merma"],
                    row["referencias_produccion"],
                ]
            )

    def _write_closure_detail_sheet(self, *, ws, closure):
        ws.append(
            [
                "receta_padre",
                "codigo_point",
                "inventario_inicial_teorico",
                "produccion_mes",
                "venta_directa_enteros",
                "venta_derivada_equivalente",
                "venta_total_equivalente",
                "merma_directa_enteros",
                "merma_derivada_equivalente",
                "merma_total_equivalente",
                "inventario_final_teorico",
                "source_sale_rows",
                "source_production_rows",
                "source_waste_rows",
                "has_catalog_issue",
                "catalog_issue_note",
            ]
        )
        for line in closure.lines.select_related("receta_padre").all():
            ws.append(
                [
                    line.receta_padre.nombre,
                    line.receta_padre.codigo_point,
                    float(line.inventario_inicial_teorico or 0),
                    float(line.produccion_mes or 0),
                    float(line.venta_directa_enteros or 0),
                    float(line.venta_derivada_equivalente or 0),
                    float(line.venta_total_equivalente or 0),
                    float(line.merma_directa_enteros or 0),
                    float(line.merma_derivada_equivalente or 0),
                    float(line.merma_total_equivalente or 0),
                    float(line.inventario_final_teorico or 0),
                    int(line.source_sale_rows or 0),
                    int(line.source_production_rows or 0),
                    int(line.source_waste_rows or 0),
                    "YES" if line.has_catalog_issue else "NO",
                    line.catalog_issue_note or "",
                ]
            )

    def _write_raw_sheet(self, *, ws, rows: list[dict[str, Any]]):
        if not rows:
            ws.append(["sin_datos"])
            return
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])

    def _write_coverage_sheet(self, *, ws, coverage_before: CoverageSnapshot, coverage_after: CoverageSnapshot):
        ws.append(["metric", "before", "after"])
        for key in [
            "sales_rows",
            "production_rows",
            "waste_rows",
            "sales_latest_date",
            "production_latest_date",
            "waste_latest_date",
        ]:
            ws.append([key, coverage_before.as_dict()[key], coverage_after.as_dict()[key]])

    def _write_sync_actions_sheet(self, *, ws, sync_actions: list[dict[str, Any]]):
        ws.append(["kind", "job_id", "status", "start_date", "end_date", "error_message", "result_summary"])
        for action in sync_actions:
            ws.append(
                [
                    action.get("kind", ""),
                    action.get("job_id"),
                    action.get("status", ""),
                    action.get("start_date", ""),
                    action.get("end_date", ""),
                    action.get("error_message", ""),
                    json.dumps(action.get("result_summary") or {}, ensure_ascii=False),
                ]
            )

    def _build_sync_action(self, *, kind: str, job, start_date: date, end_date: date) -> dict[str, Any]:
        return {
            "kind": kind,
            "job_id": getattr(job, "id", None),
            "status": getattr(job, "status", ""),
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "error_message": getattr(job, "error_message", ""),
            "result_summary": dict(getattr(job, "result_summary", {}) or {}),
        }

    def _parse_month(self, month: str | date) -> date:
        if isinstance(month, date):
            return date(month.year, month.month, 1)
        parsed = datetime.strptime(str(month).strip(), "%Y-%m").date()
        return date(parsed.year, parsed.month, 1)

    def _month_end(self, month_start: date) -> date:
        return date(month_start.year, month_start.month, monthrange(month_start.year, month_start.month)[1])

    def _date_to_text(self, value) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            if timezone.is_aware(value):
                value = timezone.localtime(value)
            return value.isoformat()
        return str(value)
