from __future__ import annotations

import json
from calendar import monthrange
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Sum
from django.utils import timezone
from openpyxl import load_workbook

from control.models import MermaMensualSucursal
from pos_bridge.models import PointDailySale, PointProductionLine, PointSalesDailyProductFact
from pos_bridge.services.sales_matching_service import PointSalesMatchingService
from recetas.models import ProductoMonthClosure, ProductoMonthClosureLine, Receta
from recetas.utils.normalizacion import normalizar_nombre
from reportes.models import FactProduccionDiaria


def _month_cursor(value: str) -> date:
    try:
        year_text, month_text = value.strip().split("-", 1)
        return date(int(year_text), int(month_text), 1)
    except Exception as exc:  # noqa: BLE001
        raise CommandError(f"Mes invalido '{value}'. Usa formato YYYY-MM.") from exc


def _parse_decimal(raw_value) -> Decimal:
    if raw_value in (None, ""):
        return Decimal("0")
    if isinstance(raw_value, Decimal):
        return raw_value
    try:
        return Decimal(str(raw_value).strip().replace(",", ""))
    except (InvalidOperation, ValueError) as exc:
        raise CommandError(f"No pude convertir a decimal el valor '{raw_value}'.") from exc


def _has_value(raw_value) -> bool:
    if raw_value in (None, ""):
        return False
    if isinstance(raw_value, str):
        return bool(raw_value.strip())
    return True


def _audit_status(*, final_difference: Decimal, total_waste: Decimal, has_physical_inventory: bool) -> str:
    if not has_physical_inventory:
        return ProductoMonthClosureLine.AUDIT_STATUS_SIN_INVENTARIO_FISICO
    if final_difference == 0:
        if total_waste > 0:
            return ProductoMonthClosureLine.AUDIT_STATUS_CUADRA_CON_MERMA
        return ProductoMonthClosureLine.AUDIT_STATUS_CUADRA
    if final_difference < 0:
        return ProductoMonthClosureLine.AUDIT_STATUS_SOBRANTE_FISICO
    return ProductoMonthClosureLine.AUDIT_STATUS_FALTANTE_NO_EXPLICADO


def _audit_detail(status: str, final_difference: Decimal) -> str:
    if status == ProductoMonthClosureLine.AUDIT_STATUS_CUADRA:
        return "El inventario teórico cuadra contra el físico."
    if status == ProductoMonthClosureLine.AUDIT_STATUS_CUADRA_CON_MERMA:
        return "El inventario cuadra considerando merma reportada."
    if status == ProductoMonthClosureLine.AUDIT_STATUS_SOBRANTE_FISICO:
        return f"Sobrante físico de {abs(final_difference)} unidades contra el teórico."
    if status == ProductoMonthClosureLine.AUDIT_STATUS_FALTANTE_NO_EXPLICADO:
        return f"Faltante no explicado de {final_difference} unidades."
    return "Sin inventario físico suficiente para auditar."


def _is_category_row(product_name: str, numeric_values: list[Decimal]) -> bool:
    normalized = normalizar_nombre(product_name)
    if not normalized or normalized == "producto":
        return False
    if product_name.strip().upper() != product_name.strip():
        return False
    return sum(1 for value in numeric_values if value != 0) <= 1


def _name_variants(raw_name: str) -> set[str]:
    normalized = normalizar_nombre(raw_name or "")
    bases = {
        normalized,
        normalized.replace(" de ", " "),
        normalized.replace(" roll ", " rol "),
        normalized.replace(" rol ", " roll "),
    }
    variants = set(bases)
    for base in bases:
        if base.endswith(" rebanada"):
            variants.add(f"{base[:-9].rstrip()} r")
        if base.endswith(" r"):
            variants.add(f"{base[:-2].rstrip()} rebanada")
    return {value.strip() for value in variants if value.strip()}


def _build_recipe_variant_map() -> dict[str, Receta]:
    result: dict[str, Receta] = {}
    for receta in Receta.objects.filter(tipo=Receta.TIPO_PRODUCTO_FINAL).order_by("id"):
        for variant in _name_variants(receta.nombre):
            result.setdefault(variant, receta)
    return result


def _resolve_receta(*, matcher: PointSalesMatchingService, variant_map: dict[str, Receta], product_name: str):
    receta = matcher.resolve_receta(codigo_point="", point_name=product_name)
    if receta is not None:
        return receta
    for variant in _name_variants(product_name):
        receta = variant_map.get(variant)
        if receta is not None:
            return receta
    return None


def _aggregate_by_recipe(queryset, field_name: str, recipe_field: str = "receta_id") -> dict[int, Decimal]:
    rows = queryset.values(recipe_field).annotate(total=Sum(field_name))
    return {
        int(row[recipe_field]): _parse_decimal(row["total"])
        for row in rows
        if row.get(recipe_field)
    }


def _sales_map(month_start: date, month_end: date) -> tuple[dict[int, Decimal], str]:
    facts = PointSalesDailyProductFact.objects.filter(
        sale_date__gte=month_start,
        sale_date__lte=month_end,
        receta_id__isnull=False,
    )
    if facts.exists():
        sales_map = _aggregate_by_recipe(facts, "total_cantidad")
        source_parts = ["PointSalesDailyProductFact"]
        for fallback_qs, field_name, source_name in [
            (
                FactProduccionDiaria.objects.filter(
                    fecha__gte=month_start,
                    fecha__lte=month_end,
                    receta_id__isnull=False,
                    vendido__gt=0,
                ),
                "vendido",
                "FactProduccionDiaria",
            ),
            (
                PointDailySale.objects.filter(
                    sale_date__gte=month_start,
                    sale_date__lte=month_end,
                    receta_id__isnull=False,
                ),
                "quantity",
                "PointDailySale",
            ),
        ]:
            fallback_map = _aggregate_by_recipe(fallback_qs, field_name)
            filled = 0
            for receta_id, value in fallback_map.items():
                if receta_id not in sales_map or sales_map[receta_id] == Decimal("0"):
                    sales_map[receta_id] = value
                    filled += 1
            if filled:
                source_parts.append(f"{source_name}({filled})")
        return sales_map, "+".join(source_parts)

    sales = PointDailySale.objects.filter(
        sale_date__gte=month_start,
        sale_date__lte=month_end,
        receta_id__isnull=False,
    )
    if sales.exists():
        return _aggregate_by_recipe(sales, "quantity"), "PointDailySale"

    fact_sales = FactProduccionDiaria.objects.filter(
        fecha__gte=month_start,
        fecha__lte=month_end,
        receta_id__isnull=False,
        vendido__gt=0,
    )
    if fact_sales.exists():
        return _aggregate_by_recipe(fact_sales, "vendido"), "FactProduccionDiaria"
    return {}, "sin_datos"


def _production_map(month_start: date, month_end: date) -> tuple[dict[int, Decimal], str]:
    facts = FactProduccionDiaria.objects.filter(
        fecha__gte=month_start,
        fecha__lte=month_end,
        receta_id__isnull=False,
    )
    if facts.exists():
        return _aggregate_by_recipe(facts, "producido"), "FactProduccionDiaria"

    point_rows = PointProductionLine.objects.filter(
        production_date__gte=month_start,
        production_date__lte=month_end,
        receta_id__isnull=False,
        is_insumo=False,
    )
    if point_rows.exists():
        return _aggregate_by_recipe(point_rows, "produced_quantity"), "PointProductionLine"
    return {}, "sin_datos"


def _merma_maps(month_start: date, month_end: date) -> tuple[dict[int, Decimal], dict[int, Decimal], str]:
    fact_rows = FactProduccionDiaria.objects.filter(
        fecha__gte=month_start,
        fecha__lte=month_end,
        receta_id__isnull=False,
        merma__gt=0,
    )
    if fact_rows.exists():
        return _aggregate_by_recipe(fact_rows, "merma"), {}, "FactProduccionDiaria"

    rows = MermaMensualSucursal.objects.filter(periodo=month_start, receta_id__isnull=False)
    if rows.exists():
        return (
            _aggregate_by_recipe(rows, "unidades_merma"),
            _aggregate_by_recipe(rows, "costo_merma"),
            "MermaMensualSucursal",
        )
    return {}, {}, "sin_datos"


class Command(BaseCommand):
    help = "Importa inventarios historicos desde el Excel operativo sin sustituir ventas, produccion ni merma."

    def add_arguments(self, parser):
        parser.add_argument("input_path", help="Ruta absoluta del Excel fuente.")
        parser.add_argument("--sheet", required=True, help="Hoja a leer, por ejemplo 'ENERO 26'.")
        parser.add_argument("--month", required=True, help="Mes del cierre en formato YYYY-MM.")
        parser.add_argument("--dry-run", action="store_true", help="Analiza y simula la importación sin persistir.")
        parser.add_argument("--rebuild", action="store_true", help="Permite reemplazar un cierre no bloqueado.")
        parser.add_argument("--actor-username", default="", help="Usuario ERP para registrar la importación.")
        parser.add_argument("--approval-note", default="", help="Nota operativa para dejar en metadata.")

    def handle(self, *args, **options):
        input_path = Path(options["input_path"]).expanduser().resolve()
        if not input_path.exists():
            raise CommandError(f"No existe el archivo '{input_path}'.")

        month_start = _month_cursor(options["month"])
        month_end = date(month_start.year, month_start.month, monthrange(month_start.year, month_start.month)[1])
        sheet_name = (options["sheet"] or "").strip()

        actor = None
        actor_username = (options.get("actor_username") or "").strip()
        if actor_username:
            actor = get_user_model().objects.filter(username=actor_username).first()
            if actor is None:
                raise CommandError(f"No existe actor_username '{actor_username}'.")

        existing = ProductoMonthClosure.objects.filter(month_start=month_start).first()
        if existing is not None and existing.is_locked:
            raise CommandError(f"El cierre {month_start:%Y-%m} esta bloqueado.")
        if existing is not None and not options.get("rebuild") and not options.get("dry_run"):
            raise CommandError(f"El cierre {month_start:%Y-%m} ya existe. Usa --rebuild si quieres reemplazarlo.")

        wb = load_workbook(filename=str(input_path), data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"La hoja '{sheet_name}' no existe en '{input_path.name}'.")
        ws = wb[sheet_name]

        matcher = PointSalesMatchingService()
        recipe_variant_map = _build_recipe_variant_map()
        rows: list[dict[str, object]] = []
        unmatched_rows: list[dict[str, object]] = []
        category = ""
        stopped_at_conversion_helper = False

        for row_idx in range(1, ws.max_row + 1):
            product_name = str(ws[f"B{row_idx}"].value or "").strip()
            if not product_name:
                continue
            normalized = normalizar_nombre(product_name)
            if normalized == "producto" and _has_value(ws[f"D{row_idx}"].value):
                stopped_at_conversion_helper = True
                break
            if normalized == "producto":
                continue

            values = {
                "inventario_inicial": _parse_decimal(ws[f"D{row_idx}"].value),
                "cedis": _parse_decimal(ws[f"I{row_idx}"].value),
                "sucursales": _parse_decimal(ws[f"J{row_idx}"].value),
                "fisico_total": _parse_decimal(ws[f"K{row_idx}"].value),
            }
            numeric_values = list(values.values())
            if _is_category_row(product_name, numeric_values):
                category = product_name.strip().title()
                continue
            if not any(value != 0 for value in numeric_values):
                continue

            receta = _resolve_receta(matcher=matcher, variant_map=recipe_variant_map, product_name=product_name)
            row_payload = {
                "row_number": row_idx,
                "category": category,
                "product_name": product_name,
                "receta": receta,
                **values,
            }
            rows.append(row_payload)
            if receta is None:
                unmatched_rows.append(
                    {
                        "row_number": row_idx,
                        "category": category,
                        "product_name": product_name,
                    }
                )

        matched_source_rows = [row for row in rows if row["receta"] is not None]
        if not matched_source_rows:
            raise CommandError("No se pudo homologar ninguna fila operativa del Excel a recetas ERP.")

        grouped_rows: dict[int, dict[str, object]] = {}
        decimal_keys = [
            "inventario_inicial",
            "cedis",
            "sucursales",
            "fisico_total",
        ]
        for source_row in matched_source_rows:
            receta = source_row["receta"]
            bucket = grouped_rows.setdefault(
                receta.id,
                {
                    "receta": receta,
                    "source_rows": [],
                    **{key: Decimal("0") for key in decimal_keys},
                },
            )
            bucket["source_rows"].append(
                {
                    "row_number": source_row["row_number"],
                    "category": source_row["category"],
                    "product_name": source_row["product_name"],
                }
            )
            for key in decimal_keys:
                bucket[key] += source_row[key]

        matched_rows = list(grouped_rows.values())
        sales_map, sales_source = _sales_map(month_start, month_end)
        production_map, production_source = _production_map(month_start, month_end)
        merma_map, merma_cost_map, merma_source = _merma_maps(month_start, month_end)

        payload = {
            "mode": "dry_run" if options.get("dry_run") else "import",
            "month": month_start.strftime("%Y-%m"),
            "sheet": sheet_name,
            "rows_read": len(rows),
            "matched_rows": len(matched_source_rows),
            "closure_lines": len(matched_rows),
            "unmatched_rows": unmatched_rows[:50],
            "stopped_at_conversion_helper": stopped_at_conversion_helper,
            "operational_sources": {
                "ventas": sales_source,
                "produccion": production_source,
                "merma": merma_source,
            },
        }
        if options.get("dry_run"):
            self.stdout.write(json.dumps(payload, ensure_ascii=False, indent=2, default=str))
            return

        with transaction.atomic():
            closure, _ = ProductoMonthClosure.objects.get_or_create(
                month_start=month_start,
                defaults={
                    "month_end": month_end,
                    "status": ProductoMonthClosure.STATUS_DRAFT,
                    "opening_source": ProductoMonthClosure.OPENING_SOURCE_BOOTSTRAP_SEED,
                },
            )
            if closure.is_locked:
                raise CommandError(f"El cierre {month_start:%Y-%m} esta bloqueado.")
            if closure.lines.exists() and not options.get("rebuild"):
                raise CommandError(f"El cierre {month_start:%Y-%m} ya tiene lineas. Usa --rebuild.")
            closure.lines.all().delete()
            closure.month_end = month_end
            closure.status = ProductoMonthClosure.STATUS_BUILT
            closure.opening_source = ProductoMonthClosure.OPENING_SOURCE_BOOTSTRAP_SEED
            closure.opening_reference_date = month_start
            closure.built_at = timezone.now()
            closure.built_by = actor
            closure.notes = options.get("approval_note") or "Importación histórica desde Excel operativo."
            closure.metadata = {
                "historical_excel_import": {
                    "source_path": str(input_path),
                    "source_file": input_path.name,
                    "source_sheet": sheet_name,
                    "scope": "inventory_only",
                    "imported_columns": [
                        "inventario_inicial",
                        "inventario_final_point_cedis",
                        "inventario_final_point_sucursales",
                        "inventario_final_point_total",
                    ],
                    "excluded_columns": [
                        "produccion_mes_desde_excel",
                        "venta_total_equivalente_desde_excel",
                        "merma_total_equivalente_desde_excel",
                    ],
                    "operational_sources": payload["operational_sources"],
                    "rows_read": len(rows),
                    "matched_rows": len(matched_source_rows),
                    "closure_lines": len(matched_rows),
                    "unmatched_rows": unmatched_rows[:50],
                    "stopped_at_conversion_helper": stopped_at_conversion_helper,
                }
            }
            closure.save()

            for row in matched_rows:
                receta = row["receta"]
                producido = production_map.get(receta.id, Decimal("0"))
                vendido = sales_map.get(receta.id, Decimal("0"))
                merma = merma_map.get(receta.id, Decimal("0"))
                inventario_teorico = row["inventario_inicial"] + producido - vendido - merma
                diferencia_inventario = inventario_teorico - row["fisico_total"]
                physical_cells_have_values = any(
                    _has_value(ws[f"{column}{source_row['row_number']}"].value)
                    for source_row in row["source_rows"]
                    for column in ["I", "J", "K"]
                )
                status = _audit_status(
                    final_difference=diferencia_inventario,
                    total_waste=merma,
                    has_physical_inventory=physical_cells_have_values,
                )
                ProductoMonthClosureLine.objects.create(
                    closure=closure,
                    receta_padre=receta,
                    inventario_inicial_teorico=row["inventario_inicial"],
                    produccion_mes=producido,
                    venta_directa_enteros=vendido,
                    venta_derivada_equivalente=Decimal("0"),
                    venta_total_equivalente=vendido,
                    merma_directa_enteros=merma,
                    merma_derivada_equivalente=Decimal("0"),
                    merma_total_equivalente=merma,
                    inventario_final_teorico=inventario_teorico,
                    inventario_final_point_cedis=row["cedis"],
                    inventario_final_point_sucursales=row["sucursales"],
                    inventario_final_point_total=row["fisico_total"],
                    diferencia_teorico_vs_point=diferencia_inventario,
                    estado_auditoria=status,
                    detalle_auditoria=_audit_detail(status, diferencia_inventario),
                    metadata={
                        "historical_excel": {
                            "scope": "inventory_only",
                            "source_rows": row["source_rows"],
                            "inventario_inicial": str(row["inventario_inicial"]),
                            "inventario_final_point_cedis": str(row["cedis"]),
                            "inventario_final_point_sucursales": str(row["sucursales"]),
                            "inventario_final_point_total": str(row["fisico_total"]),
                            "operational_sources": payload["operational_sources"],
                        }
                    },
                )

            payload["closure_id"] = closure.id
            payload["line_count"] = closure.lines.count()
        self.stdout.write(json.dumps(payload, ensure_ascii=False, indent=2, default=str))
