from __future__ import annotations

import json
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from pos_bridge.services.product_month_closure_service import ProductMonthClosureError, ProductMonthClosureService
from pos_bridge.services.sales_matching_service import PointSalesMatchingService
from recetas.utils.normalizacion import normalizar_nombre


def _month_cursor(value: str) -> date:
    try:
        year_text, month_text = value.strip().split("-", 1)
        return date(int(year_text), int(month_text), 1)
    except Exception as exc:  # noqa: BLE001
        raise CommandError(f"Mes invalido '{value}'. Usa formato YYYY-MM.") from exc


def _parse_decimal(raw_value) -> Decimal:
    if raw_value in (None, ""):
        return Decimal("0")
    if isinstance(raw_value, Decimal):
        return raw_value
    try:
        return Decimal(str(raw_value).strip().replace(",", ""))
    except (InvalidOperation, ValueError) as exc:
        raise CommandError(f"No pude convertir a decimal el valor '{raw_value}'.") from exc


def _is_structural_name(raw_name: str) -> bool:
    normalized = normalizar_nombre(raw_name or "")
    if not normalized:
        return True
    return normalized in {
        "producto",
        "productos",
        "producto rebanadas",
        "producto / rebanadas",
        "rebanadas",
    }


class Command(BaseCommand):
    help = "Siembra un cierre mensual historico desde un Excel de control para destrabar el opening del siguiente mes."

    def add_arguments(self, parser):
        parser.add_argument("input_path", help="Ruta absoluta del Excel fuente.")
        parser.add_argument("--sheet", required=True, help="Hoja a leer, por ejemplo 'SEPT 25'.")
        parser.add_argument("--seed-month", required=True, help="Mes a sembrar en formato YYYY-MM, por ejemplo 2025-08.")
        parser.add_argument("--name-column", default="B", help="Columna con el nombre del producto. Default B.")
        parser.add_argument(
            "--quantity-column",
            default="D",
            help="Columna con el inventario inicial del siguiente mes que se sembrara como final teorico. Default D.",
        )
        parser.add_argument("--dry-run", action="store_true", help="Analiza y simula la siembra sin persistir.")
        parser.add_argument("--rebuild", action="store_true", help="Permite reemplazar un cierre no bloqueado.")
        parser.add_argument("--actor-username", default="", help="Usuario ERP para registrar la ejecucion.")
        parser.add_argument("--approval-note", default="", help="Nota de aprobacion o contexto.")
        parser.add_argument("--approval-reason", default="bootstrap_excel_seed", help="Motivo corto del seed.")

    def handle(self, *args, **options):
        input_path = Path(options["input_path"]).expanduser().resolve()
        if not input_path.exists():
            raise CommandError(f"No existe el archivo '{input_path}'.")

        seed_month = _month_cursor(options["seed_month"])
        sheet_name = (options["sheet"] or "").strip()
        if not sheet_name:
            raise CommandError("--sheet es obligatorio.")

        actor = None
        actor_username = (options.get("actor_username") or "").strip()
        if actor_username:
            actor = get_user_model().objects.filter(username=actor_username).first()
            if actor is None:
                raise CommandError(f"No existe actor_username '{actor_username}'.")

        wb = load_workbook(filename=str(input_path), data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"La hoja '{sheet_name}' no existe en '{input_path.name}'.")
        ws = wb[sheet_name]

        matcher = PointSalesMatchingService()
        rows_payload: list[dict[str, object]] = []
        unmatched_rows: list[dict[str, object]] = []
        structural_rows_skipped = 0
        blank_rows_skipped = 0

        name_column = (options.get("name_column") or "B").strip().upper()
        quantity_column = (options.get("quantity_column") or "D").strip().upper()

        for row_idx in range(1, ws.max_row + 1):
            product_name = str(ws[f"{name_column}{row_idx}"].value or "").strip()
            quantity_raw = ws[f"{quantity_column}{row_idx}"].value
            if not product_name and quantity_raw in (None, ""):
                blank_rows_skipped += 1
                continue
            if _is_structural_name(product_name):
                structural_rows_skipped += 1
                continue

            quantity = _parse_decimal(quantity_raw)
            if quantity == 0:
                continue

            receta = matcher.resolve_receta(codigo_point="", point_name=product_name)
            row_payload = {
                "row_number": row_idx,
                "source_name": product_name,
                "quantity": quantity,
                "receta": receta,
            }
            rows_payload.append(row_payload)
            if receta is None:
                unmatched_rows.append(
                    {
                        "row_number": row_idx,
                        "product_name": product_name,
                        "quantity": str(quantity),
                    }
                )

        if not rows_payload:
            raise CommandError("La hoja no contiene filas operativas con cantidades distintas de cero.")

        service = ProductMonthClosureService(matcher=matcher)
        source_label = f"{input_path.name}::{sheet_name}::{quantity_column}"
        if options.get("dry_run"):
            with transaction.atomic():
                closure = service.build_bootstrap_seed(
                    month=seed_month,
                    seed_rows=rows_payload,
                    source_label=source_label,
                    source_path=str(input_path),
                    source_sheet=sheet_name,
                    built_by=actor,
                    rebuild=bool(options.get("rebuild")),
                    approval_note=options.get("approval_note") or "",
                    approval_reason=options.get("approval_reason") or "bootstrap_excel_seed",
                    approval_channel="command_bootstrap_dry_run",
                )
                metadata = dict(closure.metadata or {})
                bootstrap_seed = dict(metadata.get("bootstrap_seed") or {})
                bootstrap_seed["unmatched_products"] = [row["product_name"] for row in unmatched_rows[:50]]
                bootstrap_seed["unmatched_rows"] = unmatched_rows[:50]
                bootstrap_seed["rows_payload_count"] = len(rows_payload)
                bootstrap_seed["structural_rows_skipped"] = structural_rows_skipped
                bootstrap_seed["blank_rows_skipped"] = blank_rows_skipped
                metadata["bootstrap_seed"] = bootstrap_seed
                validation = metadata.get("validation", {})
                payload = {
                    "mode": "dry_run",
                    "month": seed_month.strftime("%Y-%m"),
                    "status": "warning" if validation.get("blocking_issues") else "ready",
                    "opening_source": closure.opening_source,
                    "line_count": closure.lines.count(),
                    "validation": validation,
                    "bootstrap_seed": bootstrap_seed,
                }
                transaction.set_rollback(True)
            self.stdout.write(json.dumps(payload, ensure_ascii=False, indent=2, default=str))
            return

        closure = service.build_bootstrap_seed(
            month=seed_month,
            seed_rows=rows_payload,
            source_label=source_label,
            source_path=str(input_path),
            source_sheet=sheet_name,
            built_by=actor,
            rebuild=bool(options.get("rebuild")),
            approval_note=options.get("approval_note") or "",
            approval_reason=options.get("approval_reason") or "bootstrap_excel_seed",
            approval_channel="command_bootstrap",
        )
        metadata = dict(closure.metadata or {})
        bootstrap_seed = dict(metadata.get("bootstrap_seed") or {})
        bootstrap_seed["unmatched_products"] = [row["product_name"] for row in unmatched_rows[:50]]
        bootstrap_seed["unmatched_rows"] = unmatched_rows[:50]
        bootstrap_seed["rows_payload_count"] = len(rows_payload)
        bootstrap_seed["structural_rows_skipped"] = structural_rows_skipped
        bootstrap_seed["blank_rows_skipped"] = blank_rows_skipped
        metadata["bootstrap_seed"] = bootstrap_seed
        closure.metadata = metadata
        closure.save(update_fields=["metadata", "updated_at"])
        payload = {
            "mode": "build",
            "month": seed_month.strftime("%Y-%m"),
            "status": closure.status,
            "opening_source": closure.opening_source,
            "line_count": closure.lines.count(),
            "validation": metadata.get("validation", {}),
            "bootstrap_seed": bootstrap_seed,
        }
        self.stdout.write(json.dumps(payload, ensure_ascii=False, indent=2, default=str))
